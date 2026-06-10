import sys
import os

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side
except ImportError:
    os.system('pip install openpyxl')
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "투입인원 명단"

ws.merge_cells('A1:F1')
ws['A1'] = "가산~가평 천연가스 공급시설 건설공사 비파괴검사 기술용역 투입인원 명단"
ws['A1'].font = Font(size=16, bold=True)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

headers = ["연번", "직책 (담당분야)", "성명", "생년월일", "서명 (인)", "비고"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col)
    cell.value = header
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 25

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

sample_data = [
    [1, "총괄 책임자", "", "", "", ""],
    [2, "방사선투과검사(RT)", "", "", "", ""],
    [3, "초음파탐상검사(UT)", "", "", "", ""],
    [4, "자기탐상검사(MT)", "", "", "", ""],
    [5, "침투탐상검사(PT)", "", "", "", ""],
]

for i in range(10):
    if i < len(sample_data):
        row_data = sample_data[i]
    else:
        row_data = [i+1, "", "", "", "", ""]
    
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=i+4, column=col)
        cell.value = val
        cell.alignment = Alignment(horizontal='center', vertical='center')

for row in ws.iter_rows(min_row=3, max_row=13, min_col=1, max_col=6):
    for cell in row:
        cell.border = thin_border
        
for r in range(3, 14):
    ws.row_dimensions[r].height = 30

excel_path = r"c:\Users\-\OneDrive\바탕 화면\home\투입인원_명단.xlsx"
wb.save(excel_path)
print(f"SUCCESS: {excel_path}")
