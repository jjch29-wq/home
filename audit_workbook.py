import json, math, re, statistics, sys
from collections import Counter, defaultdict
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils import get_column_letter

path = Path(sys.argv[1])
wf = load_workbook(path, data_only=False, read_only=False)
wv = load_workbook(path, data_only=True, read_only=False)

FORMULA_ERRS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}
sum_re = re.compile(r"^=SUM\(([^)]+)\)$", re.I)
cell_re = re.compile(r"(?:(?:'([^']+)'|([^'!]+))!)?\$?([A-Z]{1,3})\$?(\d+)")

def used_bounds(ws):
    cells = [(c.row, c.column) for row in ws.iter_rows() for c in row if c.value is not None]
    if not cells: return (1,1,1,1)
    return min(r for r,c in cells), min(c for r,c in cells), max(r for r,c in cells), max(c for r,c in cells)

def label_left(ws, row, col, width=4):
    vals=[]
    for c in range(max(1,col-width), col):
        v=ws.cell(row,c).value
        if isinstance(v,str) and not v.startswith('=') and v.strip(): vals.append(v.strip())
    return " | ".join(vals[-2:])

report={"file":str(path),"sheets":[],"formula_errors":[],"broken_refs":[],"sum_mismatches":[],"formula_gaps":[],"missing_inputs":[],"numeric_anomalies":[],"keywords":[],"workbook_links":[],"defined_names":[]}

for dn in wf.defined_names.values():
    report["defined_names"].append({"name":dn.name,"attr_text":dn.attr_text})

for ws in wf.worksheets:
    vws=wv[ws.title]
    r1,c1,r2,c2=used_bounds(ws)
    formulas=[]; errors=[]; links=[]
    # row summaries and keyword cells
    for row in ws.iter_rows(min_row=r1,max_row=r2,min_col=c1,max_col=c2):
        for c in row:
            val=c.value
            if isinstance(val,str):
                if val.startswith('='):
                    formulas.append(c.coordinate)
                    if '#REF!' in val: report['broken_refs'].append({"sheet":ws.title,"cell":c.coordinate,"formula":val})
                    if '[' in val and ']' in val: links.append({"cell":c.coordinate,"formula":val})
                else:
                    txt=val.strip()
                    if any(k in txt for k in ['직접비','간접비','공사비','원가','도급','실행','이익','일반관리','보험','안전관리','노무비','재료비','경비']):
                        report['keywords'].append({"sheet":ws.title,"cell":c.coordinate,"text":txt})
            cached=vws[c.coordinate].value
            if isinstance(cached,str) and cached in FORMULA_ERRS:
                errors.append({"cell":c.coordinate,"error":cached,"formula":val})
    report['formula_errors'] += [{"sheet":ws.title,**e} for e in errors]
    report['workbook_links'] += [{"sheet":ws.title,**e} for e in links]

    # exact SUM checks against cached numeric constituents
    for coord in formulas:
        f=ws[coord].value
        m=sum_re.match(f.replace(' ','')) if isinstance(f,str) else None
        if not m or ',' in m.group(1): continue
        target=vws[coord].value
        try:
            rng=ws[m.group(1)]
            flat=[x for row in rng for x in row] if isinstance(rng,tuple) else [rng]
            vals=[wv[ws.title][x.coordinate].value for x in flat]
            if isinstance(target,(int,float)):
                calc=sum(x for x in vals if isinstance(x,(int,float)))
                if abs(target-calc)>max(1,abs(target)*1e-8):
                    report['sum_mismatches'].append({"sheet":ws.title,"cell":coord,"formula":f,"cached":target,"recalc":calc,"delta":target-calc})
        except Exception: pass

    # gaps inside vertical/horizontal formula runs
    for col in range(c1,c2+1):
        flags=[isinstance(ws.cell(r,col).value,str) and ws.cell(r,col).value.startswith('=') for r in range(r1,r2+1)]
        for i in range(1,len(flags)-1):
            r=r1+i
            if flags[i-1] and flags[i+1] and not flags[i] and ws.cell(r,col).value is None:
                report['formula_gaps'].append({"sheet":ws.title,"cell":f"{get_column_letter(col)}{r}","axis":"vertical","above":ws.cell(r-1,col).value,"below":ws.cell(r+1,col).value})
    for row in range(r1,r2+1):
        flags=[isinstance(ws.cell(row,c).value,str) and ws.cell(row,c).value.startswith('=') for c in range(c1,c2+1)]
        for i in range(1,len(flags)-1):
            c=c1+i
            if flags[i-1] and flags[i+1] and not flags[i] and ws.cell(row,c).value is None:
                report['formula_gaps'].append({"sheet":ws.title,"cell":f"{get_column_letter(c)}{row}","axis":"horizontal","left":ws.cell(row,c-1).value,"right":ws.cell(row,c+1).value})

    # suspicious numeric values and likely qty/unit/amount patterns
    nums=[]
    for row in range(r1,r2+1):
        for col in range(c1,c2+1):
            v=vws.cell(row,col).value
            if isinstance(v,(int,float)) and not isinstance(v,bool) and math.isfinite(v): nums.append((row,col,float(v)))
    abspos=[abs(v) for _,_,v in nums if v!=0]
    if len(abspos)>=8:
        med=statistics.median(abspos)
        for row,col,v in nums:
            lab=label_left(ws,row,col)
            if v<0 and not any(k in lab for k in ['차감','할인','이익','손실','정산']):
                report['numeric_anomalies'].append({"sheet":ws.title,"cell":f"{get_column_letter(col)}{row}","type":"negative","value":v,"label":lab})
            if med and abs(v)>med*10000 and abs(v)>1e8:
                report['numeric_anomalies'].append({"sheet":ws.title,"cell":f"{get_column_letter(col)}{row}","type":"extreme_vs_sheet_median","value":v,"median":med,"label":lab})

    report['sheets'].append({"name":ws.title,"state":ws.sheet_state,"bounds":f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}","rows":r2-r1+1,"cols":c2-c1+1,"formulas":len(formulas),"formula_errors":len(errors),"merged_ranges":len(ws.merged_cells.ranges),"hidden_rows":sum(1 for r,d in ws.row_dimensions.items() if d.hidden),"hidden_cols":sum(1 for c,d in ws.column_dimensions.items() if d.hidden)})

payload=json.dumps(report,ensure_ascii=False,indent=2,default=str)
if len(sys.argv)>2:
    Path(sys.argv[2]).write_text(payload,encoding='utf-8')
else:
    print(payload)
