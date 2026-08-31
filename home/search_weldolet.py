import openpyxl

file_path = r'C:\Users\-\OneDrive\바탕 화면\산출내역서(가산~가평 천연가스 공급시설 건설공사 비파괴검사 기술용역)_감독경유.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

keywords = ['weldolet', '웰더렛', '웰도렛', 'wed', '4(wed)', '4"(wed)']

print("Searching for weldolet related terms...")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r in range(1, ws.max_row + 1):
        row_vals = []
        match_found = False
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r, column=c).value)
            row_vals.append(val if val != 'None' else "")
            if any(k.lower() in val.lower() for k in keywords):
                match_found = True
        
        if match_found:
            # Print non-empty columns up to max 15
            filtered_vals = [v for v in row_vals[:15] if v]
            if filtered_vals:
                print(f"[{sheet_name}] Row {r}: " + " | ".join(filtered_vals))
