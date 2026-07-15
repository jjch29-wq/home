import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

for r in range(1, 10):
    ws.cell(row=r, column=1, value=f"Data {r}")
    ws.row_dimensions[r].height = 20.55

# Apply custom dimensions
for r in range(1, 6):
    ws.row_dimensions[r].height = 40.0
    ws.row_dimensions[r].customHeight = True

wb.save("test_height.xlsx")

wb2 = openpyxl.load_workbook("test_height.xlsx")
ws2 = wb2.active
print(f"Row 1 height: {ws2.row_dimensions[1].height}")
print(f"Row 6 height: {ws2.row_dimensions[6].height}")
