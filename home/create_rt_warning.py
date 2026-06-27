import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

def create_warning_sign(output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RT Room 주의사항"

    # A4 세로 인쇄 설정
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    
    # 여백 최소화
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    ws.print_area = 'A1:A6'

    # 열 너비 (A4 가로 폭 꽉 차게)
    ws.column_dimensions['A'].width = 110

    # 행 높이 (총 800~900 사이로 배분하여 A4 세로 꽉 차게)
    ws.row_dimensions[1].height = 180  # 타이틀
    ws.row_dimensions[2].height = 80   # 서브타이틀
    ws.row_dimensions[3].height = 180  # 뱀 주의
    ws.row_dimensions[4].height = 180  # 야생동물
    ws.row_dimensions[5].height = 80   # 여백용
    ws.row_dimensions[6].height = 150  # 신고처

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thick = Side(style='thick', color='000000')
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    # 전체 테두리 및 배경색 설정을 위한 헬퍼
    def set_cell(row, val, size, color='000000', bold=True, bg_color=None):
        c = ws.cell(row=row, column=1)
        c.value = val
        c.font = Font(name='맑은 고딕', size=size, bold=bold, color=color)
        c.alignment = align_center
        c.border = border_thick
        if bg_color:
            c.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        return c

    # A1: Title (노란 배경, 빨간 글씨로 경고 강조)
    set_cell(1, "RT Room 내부 주의 사항", 65, color='FF0000', bg_color='FFFF00')

    # A2: Subtitle
    set_cell(2, "※ 촬영 前 주변 확인 사항 ※", 35, bg_color='F2F2F2')

    # A3: 뱀 주의
    set_cell(3, "🐍 뱀 주 의 🐍\n(배관 내부 및 아래 확인)", 55)

    # A4: 야생동물 포획 금지
    set_cell(4, "🐾 야생 동물 포획 금지 🐾\n발견 시 포획하지 말고 즉시 신고", 50)

    # A5: 시각적 분리 여백
    set_cell(5, "", 10)

    # A6: 신고처 (파란색 강조)
    set_cell(6, "📞 신 고 처\n031-8030-4451 (연천)", 55, color='0000FF', bg_color='E6E6FA')

    # 파일 저장
    wb.save(output_path)
    print(f"안내문이 바탕화면에 저장되었습니다: {output_path}")


if __name__ == "__main__":
    # 바탕화면 경로 탐색 (OneDrive 환경 포함)
    desktop = os.path.join(os.path.expanduser("~"), "OneDrive", "바탕 화면")
    if not os.path.exists(desktop):
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        
    output_file = os.path.join(desktop, "RT_Room_주의사항_안내문.xlsx")
    
    # 엑셀 파일 생성 실행
    create_warning_sign(output_file)
