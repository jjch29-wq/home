import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

output_path = r"C:\Users\-\OneDrive\바탕 화면\4.4.1_위험성평가표(PT)_최종.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "PT 위험성평가표"

# Styles
bold_font = Font(bold=True)
title_font = Font(bold=True, size=18)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

def set_border(ws, min_col, min_row, max_col, max_row):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border

for i in range(1, 23):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 4

# --- Header Section ---
ws.merge_cells('A1:U1')
ws['A1'] = "합동 위험성평가표 [ ■최초, □정기, □수시 ]"
ws['A1'].font = title_font
ws['A1'].alignment = center_align
ws.row_dimensions[1].height = 40

ws.merge_cells('A2:B3')
ws['A2'] = "작 업 명"
ws.merge_cells('C2:E3')
ws['C2'] = "가산~가평 천연가스 공급설비\n건설공사 비파괴검사 기술용역"
ws.merge_cells('F2:G2')
ws['F2'] = "평가기간"
ws.merge_cells('H2:M2')
ws['H2'] = "2026. 07. 09"
ws.merge_cells('N2:P2')
ws['N2'] = "재평가일"
ws.merge_cells('Q2:U2')

ws.merge_cells('A4:B5')
ws['A4'] = "회 사 명"
ws.merge_cells('C4:E5')
ws['C4'] = "서울검사(주)"

ws.merge_cells('A6:B7')
ws['A6'] = "작업공종"
ws.merge_cells('C6:E7')
ws['C6'] = "침투탐상검사"

ws.merge_cells('A8:B9')
ws['A8'] = "작업기간"
ws.merge_cells('C8:E9')
ws['C8'] = "2026.06.20 ~ 2026.06.30"

ws.merge_cells('F3:F9')
ws['F3'] = "작업전"

ws.merge_cells('G3:G5')
ws['G3'] = "수급인"

# Equal widths for signatures: H~I (2), J~K (2), L~M (2)
ws.merge_cells('H3:I3')
ws['H3'] = "근 로 자"
ws.merge_cells('J3:K3')
ws['J3'] = "작 성 자"
ws.merge_cells('L3:M3')
ws['L3'] = "승 인 자"

ws.merge_cells('H4:I5')
ws['H4'] = "유상훈 (서명)"
ws.merge_cells('J4:K5')
ws['J4'] = "주진철 (서명)"
ws.merge_cells('L4:M5')
ws['L4'] = "강신태 (서명)"

ws.merge_cells('G6:G9')
ws['G6'] = "도급인"
ws.merge_cells('H6:I6')
ws['H6'] = "검토자(감독)"
ws.merge_cells('J6:K6')
ws['J6'] = "검토자(안전)"
ws.merge_cells('L6:M6')
ws['L6'] = "승 인 자"
ws.merge_cells('H7:I9')
ws['H7'] = "(서명)"
ws.merge_cells('J7:K9')
ws['J7'] = "(서명)"
ws.merge_cells('L7:M9')
ws['L7'] = "(서명)"

ws.merge_cells('N3:N9')
ws['N3'] = "재평가"

ws.merge_cells('O3:O5')
ws['O3'] = "수급인"

ws.merge_cells('P3:Q3')
ws['P3'] = "근 로 자"
ws.merge_cells('R3:S3')
ws['R3'] = "작 성 자"
ws.merge_cells('T3:U3')
ws['T3'] = "승 인 자"

ws.merge_cells('P4:Q5')
ws['P4'] = "(서명)"
ws.merge_cells('R4:S5')
ws['R4'] = "(서명)"
ws.merge_cells('T4:U5')
ws['T4'] = "(서명)"

ws.merge_cells('O6:O9')
ws['O6'] = "도급인"
ws.merge_cells('P6:Q6')
ws['P6'] = "검토자(감독)"
ws.merge_cells('R6:S6')
ws['R6'] = "검토자(안전)"
ws.merge_cells('T6:U6')
ws['T6'] = "승 인 자"
ws.merge_cells('P7:Q9')
ws['P7'] = "(서명)"
ws.merge_cells('R7:S9')
ws['R7'] = "(서명)"
ws.merge_cells('T7:U9')
ws['T7'] = "(서명)"

for r in range(2, 10):
    for c in range(1, 22):
        cell = ws.cell(row=r, column=c)
        cell.alignment = center_align
        cell.font = Font(size=10)

set_border(ws, 1, 2, 21, 9)

# --- Table Sub-Header ---
ws.merge_cells('A10:M10')
ws['A10'] = "작업 전 평가"
ws.merge_cells('N10:U10')
ws['N10'] = "재평가"

ws.merge_cells('A11:A12')
ws['A11'] = "세부작업"
ws.merge_cells('B11:E11')
ws['B11'] = "유해 위험 요인 파악"
ws.merge_cells('F11:G12')
ws['F11'] = "현재의 안전 보건 조치"
ws.merge_cells('H11:J11')
ws['H11'] = "위험성 결정"
ws.merge_cells('K11:O12')
ws['K11'] = "위험성 감소 대책"
ws.merge_cells('P11:P12')
ws['P11'] = "개선\n예정일"
ws.merge_cells('Q11:Q12')
ws['Q11'] = "담당자\n(담당부서)"
ws.merge_cells('R11:R12')
ws['R11'] = "완료일"
ws.merge_cells('S11:U11')
ws['S11'] = "개선 후 위험성 (재평가)"

ws.merge_cells('B12:B12')
ws['B12'] = "위험\n분류"
ws.merge_cells('C12:E12')
ws['C12'] = "위험 발생 상황 및 결과"
ws['H12'] = "가능성\n(빈도)"
ws['I12'] = "중대성\n(강도)"
ws['J12'] = "위험성\n(등급)"
ws['S12'] = "가능성\n(빈도)"
ws['T12'] = "중대성\n(강도)"
ws['U12'] = "위험성\n(등급)"

for r in range(10, 13):
    for c in range(1, 22):
        cell = ws.cell(row=r, column=c)
        cell.alignment = center_align
        cell.font = bold_font
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

set_border(ws, 1, 10, 21, 12)

# Column Widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 6
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 7 
ws.column_dimensions['I'].width = 7 
ws.column_dimensions['J'].width = 7 
ws.column_dimensions['K'].width = 7 
ws.column_dimensions['L'].width = 7 
ws.column_dimensions['M'].width = 7 

ws.column_dimensions['O'].width = 25 
ws.column_dimensions['P'].width = 7
ws.column_dimensions['Q'].width = 7
ws.column_dimensions['R'].width = 7

# --- Data Definition (PT specifically mapped from screenshots) ---
data = [
    # 1. 사전준비
    ["사전준비", "기계적\n(설비)\n요인", "지하 작업 중 낙하물에 의한 사고위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 굴착 상단부에 자재 및 공구 적재 금지\n2. 하부 작업 시 상부 타 공정 혼재 작업 원칙적 금지"],
    ["사전준비", "기계적\n(설비)\n요인", "손상된 슬링벨트 사용 시 판단에 의한 중량물낙하 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 작업 전 슬링벨트 손상 및 규격 확인\n2. 파손된 슬링벨트는 현장에서 즉시 파기 및 교체"],
    ["사전준비", "기계적\n(설비)\n요인", "중량물 줄걸이작업 부적합 시 낙하에 의한 사고위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 지정된 줄걸이 용구 사용 및 결속 상태 교차 점검\n2. 관리감독자(또는 신호수) 지휘 하에 안전하게 인양 진행"],
    ["사전준비", "기계적\n(설비)\n요인", "근로자가 안전모 등 개인 보호구 미착용에 의한 충돌", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 작업구역 진입 전 개인보호구 착용 상태 점검\n2. 안전모 턱끈 결속 철저"],
    ["사전준비", "기계적\n(설비)\n요인", "중장비 이용 중량물 취급 작업 중 신호오류에 의한 충돌사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 장비 작업 반경 내 출입 통제 및 펜스 설치\n2. 전담 신호수 배치 및 운전원 간 무전 소통 확립"],

    # 2. 비파괴 검사
    ["비파괴\n검사", "화학(물질)적\n요인", "대상 화학물질에 대한 유해위험성 미인식에 의한 건강장해 발생 위험", "MSDS 준수 보호구 착용", 1, 1, "VI", "1. 침투탐상제(세척액, 침투액, 현상액) MSDS 비치\n2. 작업 전 특별 안전보건 교육 실시 및 취급 주의 전파"],
    ["비파괴\n검사", "화학(물질)적\n요인", "화학물질 취급작업시 보호구 미착용으로 인한 사고 위험", "MSDS 준수 보호구 착용", 1, 1, "VI", "1. 내화학 장갑 및 방독마스크 지급 및 착용 의무화\n2. 보호구 훼손 시 즉시 새 제품으로 교체 지급"],
    ["비파괴\n검사", "작업특성\n요인", "협소공간 검사 시 부자연스런 자세에 의한 근골격계질환 위험", "작업 전 스트레칭", 2, 1, "VI", "1. 작업 전·후 전신 스트레칭 실시\n2. 장시간 구부린 자세 유지 시 주기적인 휴식 부여"],
    ["비파괴\n검사", "작업환경\n요인", "우천으로 인한 굴착법면상태 악화에 의한 붕괴 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 작업 전 굴착 법면 균열, 용수(물비침) 일일 점검\n2. 호우 경보 시 굴착 지점 내부 진입 전면 금지"],

    # 3. 작업후 정리
    ["작업후\n정리", "기계적(설비)적\n요인", "중량물 줄걸이 작업 부적합 시 낙하에 의한 사고위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 자재 반출 시에도 2줄 걸이 이상 결속 원칙 준수\n2. 인양물 하부 절대 접근 금지"],
    ["작업후\n정리", "기계적(설비)적\n요인", "손상된 슬링벨트 사용 시 판단에 의한 중량물낙하 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 철수 작업 전 슬링벨트 훼손 유무 재차 확인\n2. 손상 의심 시 무리한 사용 금지 및 즉각 폐기"],
    ["작업후\n정리", "기계적(설비)적\n요인", "중장비 이용 중량물 취급 작업 중 신호오류에 의한 충돌사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 장비 철수 동선 사전 파악 및 신호수/유도자 통제\n2. 운전원과 수신호 및 무전 교신 확인 후 이동"]
]

start_row = 13
current_row = 13

groups = {}
for item in data:
    cat = item[0]
    groups[cat] = groups.get(cat, 0) + 1

for item in data:
    ws.cell(row=current_row, column=1, value=item[0]).alignment = center_align
    ws.cell(row=current_row, column=2, value=item[1]).alignment = center_align
    
    ws.merge_cells(f'C{current_row}:E{current_row}')
    ws.cell(row=current_row, column=3, value=item[2]).alignment = left_align
    ws.merge_cells(f'F{current_row}:G{current_row}')
    ws.cell(row=current_row, column=6, value=item[3]).alignment = left_align
    ws.cell(row=current_row, column=8, value=item[4]).alignment = center_align
    ws.cell(row=current_row, column=9, value=item[5]).alignment = center_align
    ws.cell(row=current_row, column=10, value=item[6]).alignment = center_align
    
    ws.merge_cells(f'K{current_row}:O{current_row}')
    ws.cell(row=current_row, column=11, value=item[7]).alignment = left_align
    
    ws.row_dimensions[current_row].height = 40
    current_row += 1

r = 13
for cat, count in groups.items():
    if count > 1:
        ws.merge_cells(f'A{r}:A{r+count-1}')
    r += count

def merge_sub_categories():
    r_idx = 13
    while r_idx < current_row:
        val = ws.cell(row=r_idx, column=2).value
        count = 1
        for j in range(r_idx + 1, current_row):
            if ws.cell(row=j, column=2).value == val and ws.cell(row=j, column=1).value == ws.cell(row=r_idx, column=1).value:
                count += 1
            else:
                break
        if count > 1:
            ws.merge_cells(f'B{r_idx}:B{r_idx+count-1}')
        r_idx += count

merge_sub_categories()

set_border(ws, 1, 13, 21, current_row - 1)

try:
    wb.save(output_path)
    print(f"Excel file created at: {output_path}")
except PermissionError:
    alt_path = output_path.replace(".xlsx", "_1.xlsx")
    wb.save(alt_path)
    print(f"Excel file created at: {alt_path}")
