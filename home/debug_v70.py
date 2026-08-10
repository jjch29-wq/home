import os
import glob
import openpyxl

paths = [
    r"C:\Users\jjch2\Desktop\*.xlsx",
    r"C:\Users\-\Desktop\*.xlsx"
]

found = False
for p in paths:
    files = glob.glob(p)
    for f in files:
        if 'V70' in f:
            print("Found:", f)
            try:
                wb = openpyxl.load_workbook(f, data_only=True)
                print(f"Sheets: {len(wb.sheetnames)}")
                for i, sn in enumerate(wb.sheetnames):
                    print(f"  {i}: {sn.encode('utf-8')}")
                found = True
            except Exception as e:
                print("Error:", e)
if not found:
    print("Could not find V70 template.")
