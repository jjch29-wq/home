import openpyxl, os
base = os.path.dirname(os.path.abspath(__file__))
path = os.path.join(base, 'data', 'Report_Template_현장사용량.xlsx')
wb = openpyxl.load_workbook(path)
sheet = wb.active

with open(os.path.join(base, 'data', 'columns.txt'), 'w', encoding='utf-8') as f:
    for row in [29, 30]:
        f.write(f"--- ROW {row} ---\n")
        for col in range(2, 20):
            cell = sheet.cell(row=row, column=col)
            val = cell.value
            if isinstance(cell, openpyxl.cell.cell.MergedCell) or val is None:
                for mr in sheet.merged_cells.ranges:
                    if sheet.cell(row=row, column=col).coordinate in mr:
                        val = sheet.cell(row=mr.min_row, column=mr.min_col).value
                        break
            if val and str(val).strip():
                f.write(f"Col {col}: {repr(val)}\n")
