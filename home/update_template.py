import openpyxl

template_path = r'C:\Users\-\OneDrive\바탕 화면\템플릿 가산가평월간용역진도보고서.xlsx'
wb = openpyxl.load_workbook(template_path)

project_name = "가산~가평 천연가스 공급시설 건설공사 비파괴검사 기술용역"
project_name_with_newline = "가산~가평 천연가스 공급시설 건설공사\n비파괴검사 기술용역"
prefix = "용역명 : "

updates = [
    ('표지', 8, 1, project_name_with_newline),
    ('2. 공정율 (전체)', 3, 1, prefix + project_name),
    ('2. 공정율 (주배관)', 3, 1, prefix + project_name),
    ('2. 공정율 (관리소)', 3, 1, prefix + project_name),
    ('6.3 인원현황 누계', 2, 1, prefix + project_name),
    ('8. 안전', 2, 1, prefix + project_name)
]

for sheet_name, r, c, val in updates:
    if sheet_name in wb.sheetnames:
        wb[sheet_name].cell(row=r, column=c).value = val
        print(f"Updated {sheet_name} R{r}C{c} -> {val}")

wb.save(template_path)
print("Template updated successfully.")
