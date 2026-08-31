import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\-\OneDrive\바탕 화면\템플릿 가산가평월간용역진도보고서.xlsx', data_only=True)
ws = wb['2. 공정율 (전체)']
for r in range(8, 25):
    c1 = ws.cell(row=r, column=1).value
    c2 = ws.cell(row=r, column=2).value
    print(f"Row {r}: {c1} | {c2}")
