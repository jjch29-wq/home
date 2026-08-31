import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\-\OneDrive\바탕 화면\템플릿 가산가평월간용역진도보고서.xlsx', data_only=True)
if '4. 물량세부내역 (주배관)' in wb.sheetnames:
    ws = wb['4. 물량세부내역 (주배관)']
    headers = []
    for col in range(1, 25):
        val4 = str(ws.cell(row=4, column=col).value or '').strip()
        val5 = str(ws.cell(row=5, column=col).value or '').strip()
        headers.append(f"Col {col}: {val4} / {val5}")
    with open('headers4.txt', 'w', encoding='utf-8') as f:
        f.write("\n".join(headers))
else:
    with open('headers4.txt', 'w', encoding='utf-8') as f:
        f.write("Sheet not found")
