import openpyxl

filepath = r'C:\Users\-\OneDrive\바탕 화면\지역난방 PAUT.xlsx'
try:
    wb = openpyxl.load_workbook(filepath)
    sheet_name = next((sn for sn in wb.sheetnames if '을지' in sn or 'DATA' in sn.upper()), wb.sheetnames[0])
    ws = wb[sheet_name]
    
    # Change row 17 to 20 height
    for r in range(17, 21):
        ws.row_dimensions[r].height = 50.0
        
    wb.save(r'C:\Users\-\OneDrive\바탕 화면\TEST_PAUT.xlsx')
    print("Test file saved.")
except Exception as e:
    print(f"Error: {e}")
