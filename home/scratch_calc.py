import openpyxl

filepath = r'C:\Users\-\OneDrive\바탕 화면\지역난방 PAUT.xlsx'
try:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    sheet_name = None
    for sn in wb.sheetnames:
        if '을지' in sn or 'DATA' in sn.upper() or 'Data' in sn:
            sheet_name = sn
            break
            
    if not sheet_name:
        sheet_name = wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]

    ws = wb[sheet_name]
    
    col_start = 'N'
    col_end = 'X'
    
    s = openpyxl.utils.column_index_from_string(col_start)
    e = openpyxl.utils.column_index_from_string(col_end)
    
    tot = 0.0
    w_list = []
    
    for i in range(s, e + 1):
        let = openpyxl.utils.get_column_letter(i)
        cd = ws.column_dimensions.get(let)
        w = cd.width if cd and cd.width else ws.sheet_format.defaultColWidth
        if w is None:
            w = 8.38
        w_list.append(f"{let}:{w}")
        tot += w
        
    print(f"Sheet: {sheet_name}")
    print(f"Widths: {', '.join(w_list)}")
    print(f"Total: {tot}")
except Exception as e:
    print(f"Error: {e}")
