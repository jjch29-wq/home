import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\-\OneDrive\바탕 화면\템플릿 가산가평월간용역진도보고서.xlsx', data_only=True)
target_sheets = [s for s in wb.sheetnames if '3. 비파괴검사' in s]
print(f"Target sheets: {target_sheets}")

for sheet_name in target_sheets:
    ws = wb[sheet_name]
    print(f"\n--- {sheet_name} ---")
    for r in range(1, 15):
        row_vals = []
        for c in range(1, 10):
            val = ws.cell(row=r, column=c).value
            row_vals.append(str(val) if val is not None else "")
        print(f"Row {r}: " + " | ".join(row_vals))
