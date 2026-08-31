import openpyxl

template_path = r'C:\Users\-\OneDrive\바탕 화면\템플릿 가산가평월간용역진도보고서.xlsx'
wb = openpyxl.load_workbook(template_path, data_only=True)

sheets_to_check = [
    '5.1 월별 용접불량(2026년)',
    '5.1 월별 용접불량(2027년)',
    '5.1 26년~27년 용접불량 현황',
    '5.2 용접사 불량률-주배관',
    '5.2 용접사 불량률-관리소'
]

for sheet_name in sheets_to_check:
    if sheet_name in wb.sheetnames:
        print(f"\n--- {sheet_name} ---")
        ws = wb[sheet_name]
        for r in range(1, 20):
            row_vals = []
            for c in range(1, 15):
                val = str(ws.cell(row=r, column=c).value)
                row_vals.append(val if val != 'None' else "")
            if any(row_vals):
                print(f"Row {r}: " + " | ".join(row_vals))
    else:
        print(f"Sheet not found: {sheet_name}")
