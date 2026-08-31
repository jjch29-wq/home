import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\-\OneDrive\바탕 화면\템플릿 가산가평월간용역진도보고서.xlsx', data_only=True)
lines = []
lines.append(f'Sheets: {wb.sheetnames}')

for ws in wb.worksheets:
    lines.append(f'\nSheet: {ws.title}')
    for r in range(1, 40):
        row_vals = []
        for c in range(1, 20):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                row_vals.append(f"Col {c}: {str(val).strip()}")
        if row_vals:
            lines.append(f'  Row {r}: ' + ', '.join(row_vals))

with open('c:/Users/-/PMI/home/excel_structure.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
