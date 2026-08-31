import openpyxl

excel_path = r"C:\Users\-\OneDrive\바탕 화면\템플릿 가산가평월간용역진도보고서.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

for sheet_name in ["나. 검사물량 세부내역 (주배관)", "나. 검사물량 세부내역 (관리소)"]:
    print(f"--- {sheet_name} ---")
    ws = wb[sheet_name]
    for r in range(20, 100):
        val1 = ws.cell(row=r, column=1).value
        if isinstance(val1, str):
            if "초음파탐상" in val1 or "UT" in val1.upper():
                print(f"Row {r}: found UT section -> {val1}")
            if "액체침투탐상" in val1 or "PT" in val1.upper():
                print(f"Row {r}: found PT section -> {val1}")
            if "자분탐상" in val1 or "MT" in val1.upper():
                print(f"Row {r}: found MT section -> {val1}")
            
            if "총" in val1 and "계" in val1 or "합" in val1 and "계" in val1:
                # print the sum
                vals = [str(ws.cell(row=r, column=c).value) for c in range(2, 9)]
                print(f"  Row {r} (Sum): {vals}")
