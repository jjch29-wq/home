import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\-\OneDrive\바탕 화면\템플릿 가산가평월간용역진도보고서.xlsx', data_only=True)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r in range(1, 10):
        for c in range(1, 10):
            val = str(ws.cell(row=r, column=c).value)
            if "용역명" in val or "공사명" in val or "가산" in val:
                print(f"[{sheet}] R{r}C{c}: {val}")
