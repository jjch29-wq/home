import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
import glob
from datetime import datetime
import re
import sys
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont

CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "personnel_equipment_db.json")
TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")

def calculate_experience(date_string):
    if not date_string or date_string.strip() == "-":
        return "-"
    # Extract all dates in format YYYY.MM.DD
    dates = re.findall(r"(\d{4})\.(\d{2})\.(\d{2})", date_string)
    if not dates:
        return "-"
    
    earliest_date = min(datetime(int(y), int(m), int(d)) for y, m, d in dates)
    now = datetime.now()
    
    diff_months = (now.year - earliest_date.year) * 12 + now.month - earliest_date.month
    if now.day < earliest_date.day:
        diff_months -= 1
        
    years = diff_months // 12
    months = diff_months % 12
    
    if years > 0 and months > 0:
        return f"{years}년 {months}개월"
    elif years > 0:
        return f"{years}년"
    elif months > 0:
        return f"{months}개월"
    else:
        return "1개월 미만"

class DeploymentApp:
    def __init__(self, root):
        self.root = root
        self.root.title("인원 및 장비투입계획서 자동 생성기")
        # 과도하게 크지 않은 적당한 기본 창 크기로 조절
        self.root.geometry("750x750")
        
        self.load_db()
        self.templates = glob.glob(os.path.join(TEMPLATE_DIR, "양식_*.xlsx"))
        
        self.setup_ui()
        
    def load_db(self):
        self.db = {"personnel": [], "equipment": []}
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    self.db = json.load(f)
            except:
                pass
        self.personnel_names = [p["name"] for p in self.db["personnel"]]
        
    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.pack(fill='both', expand=True)
        
        ttk.Label(main_frame, text="인원 및 장비투입계획서 생성", font=('Malgun Gothic', 16, 'bold')).pack(pady=(0, 10))
        
        # Template selection
        tmpl_frame = ttk.LabelFrame(main_frame, text="양식(템플릿) 선택", padding=10)
        tmpl_frame.pack(fill='x', pady=5)
        self.tmpl_var = tk.StringVar()
        tmpl_cb = ttk.Combobox(tmpl_frame, textvariable=self.tmpl_var, state='readonly', width=50)
        tmpl_cb['values'] = [os.path.basename(t) for t in self.templates]
        if tmpl_cb['values']:
            tmpl_cb.current(0)
        tmpl_cb.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(tmpl_frame, text="DB 관리 (인원/장비)", command=self.manage_db).pack(side=tk.RIGHT, padx=5)
        
        # Notebook for Tabs
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill='both', expand=True, pady=5)
        
        # --- TAB 1: Personnel ---
        pers_tab = ttk.Frame(notebook)
        notebook.add(pers_tab, text="인력 배치")
        
        pers_frame = ttk.Frame(pers_tab, padding=10)
        # 내용이 좌측 상단에 뭉치지 않도록 가운데(위) 정렬
        pers_frame.pack(anchor='n', pady=20)
        
        self.personnel_vars = []
        self.role_vars = []
        self.personnel_comboboxes = []
        
        headers = ["순번", "성명", "담당업무"]
        for i, h in enumerate(headers):
            ttk.Label(pers_frame, text=h, font=('Malgun Gothic', 10, 'bold')).grid(row=0, column=i, padx=15, pady=5)
            
        for i in range(15):
            ttk.Label(pers_frame, text=str(i+1)).grid(row=i+1, column=0, padx=15, pady=2)
            p_var = tk.StringVar()
            cb = ttk.Combobox(pers_frame, textvariable=p_var, state='readonly', values=[""] + self.personnel_names, width=20)
            cb.grid(row=i+1, column=1, padx=15, pady=2)
            self.personnel_vars.append(p_var)
            self.personnel_comboboxes.append(cb)
            
            r_var = tk.StringVar()
            rcb = ttk.Combobox(pers_frame, textvariable=r_var, width=30)
            rcb['values'] = ["", "현장대리인", "방사선안전관리자", "RT 팀장", "PAUT 팀장", "PT 팀장", "MT 팀장", "RT 검사자", "PAUT 검사자", "PT 검사자", "MT 검사자", "MT/PT 검사자", "PAUT 검사보조"]
            rcb.grid(row=i+1, column=2, padx=15, pady=2)
            self.role_vars.append(r_var)
            
        if len(self.personnel_vars) > 0: self.personnel_vars[0].set("주진철"); self.role_vars[0].set("현장대리인")
        if len(self.personnel_vars) > 1: self.personnel_vars[1].set("진병학"); self.role_vars[1].set("방사선안전관리자")
        if len(self.personnel_vars) > 2: self.personnel_vars[2].set("김춘호"); self.role_vars[2].set("RT 팀장")
        
        # --- TAB 2: Equipment ---
        equip_tab = ttk.Frame(notebook)
        notebook.add(equip_tab, text="장비 배치")
        
        # Create scrollable canvas for equipment
        canvas = tk.Canvas(equip_tab)
        scrollbar = ttk.Scrollbar(equip_tab, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 장비 탭도 좌측에 너무 쏠리지 않도록 캔버스의 너비를 여유있게 배치
        canvas.pack(side="left", fill="both", expand=True, padx=20, pady=20)
        scrollbar.pack(side="right", fill="y")
        
        headers_eq = ["순번", "불러오기(DB)", "수량", "구분", "품명", "규격"]
        for i, h in enumerate(headers_eq):
            ttk.Label(scrollable_frame, text=h, font=('Malgun Gothic', 10, 'bold')).grid(row=0, column=i, padx=15, pady=5)
            
        self.equip_vars = []
        
        # Build equipment map for combobox
        eq_list = self.db.get("equipment", [])
        self.eq_map = {f"[{eq.get('category', '')}] {eq.get('name', '')}": eq for eq in eq_list}
        eq_options = ["[비우기]"] + list(self.eq_map.keys())
        
        for i in range(15):
            ttk.Label(scrollable_frame, text=str(i+1)).grid(row=i+1, column=0, padx=15, pady=2)
            
            db_var = tk.StringVar()
            cb_db = ttk.Combobox(scrollable_frame, textvariable=db_var, state='readonly', values=eq_options, width=20)
            cb_db.grid(row=i+1, column=1, padx=5, pady=2)
            
            qty_var = tk.StringVar()
            cat_var = tk.StringVar()
            name_var = tk.StringVar()
            spec_var = tk.StringVar()
            
            ent_q = ttk.Entry(scrollable_frame, textvariable=qty_var, width=8)
            ent_q.grid(row=i+1, column=2, padx=5, pady=2)
            
            ent_c = ttk.Entry(scrollable_frame, textvariable=cat_var, width=15)
            ent_c.grid(row=i+1, column=3, padx=5, pady=2)
            
            ent_n = ttk.Entry(scrollable_frame, textvariable=name_var, width=25)
            ent_n.grid(row=i+1, column=4, padx=5, pady=2)
            
            ent_s = ttk.Entry(scrollable_frame, textvariable=spec_var, width=30)
            ent_s.grid(row=i+1, column=5, padx=5, pady=2)
            
            def on_eq_select(event, q_var=qty_var, c_var=cat_var, n_var=name_var, s_var=spec_var, cb=cb_db):
                selected = cb.get()
                if selected in self.eq_map:
                    eq = self.eq_map[selected]
                    q_val = str(eq.get("qty", ""))
                    if q_val == "0": q_val = ""
                    q_var.set(q_val)
                    c_var.set(eq.get("category", ""))
                    n_var.set(eq.get("name", ""))
                    s_var.set(eq.get("spec", ""))
                elif selected == "[비우기]" or not selected:
                    cb.set("")  # 콤보박스 텍스트도 완전히 비워줍니다.
                    q_var.set("")
                    c_var.set("")
                    n_var.set("")
                    s_var.set("")
                    
            cb_db.bind("<<ComboboxSelected>>", on_eq_select)
            
            self.equip_vars.append({
                "db_var": db_var,
                "qty_var": qty_var,
                "cat_var": cat_var,
                "name_var": name_var,
                "spec_var": spec_var
            })
            
        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=(5, 10))
        
        # 생성 버튼을 크고 명확하게 정중앙 하단에 배치
        ttk.Button(btn_frame, text="엑셀 생성하기", command=self.generate_excel, style='Accent.TButton').pack(side=tk.BOTTOM, ipadx=40, ipady=15)
        
        self.load_session()
        
    def manage_db(self):
        top = tk.Toplevel(self.root)
        top.title("데이터베이스 관리")
        top.geometry("700x600")
        
        notebook = ttk.Notebook(top)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # --- Personnel Tab ---
        p_tab = ttk.Frame(notebook)
        notebook.add(p_tab, text="인력 관리")
        
        columns = ("name", "qual", "date")
        tree = ttk.Treeview(p_tab, columns=columns, show="headings", height=15)
        tree.heading("name", text="이름")
        tree.heading("qual", text="자격사항")
        tree.heading("date", text="취득일(형식:YYYY.MM.DD(RT))")
        tree.column("name", width=80)
        tree.column("qual", width=200)
        tree.column("date", width=250)
        tree.pack(fill='both', expand=True, padx=10, pady=10)
        
        def refresh_tree():
            for item in tree.get_children(): tree.delete(item)
            for p in self.db["personnel"]: tree.insert("", "end", values=(p["name"], p.get("qualifications", ""), p.get("date", "")))
        refresh_tree()
        
        p_input = ttk.Frame(p_tab)
        p_input.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(p_input, text="이름:").grid(row=0, column=0, padx=5)
        name_ent = ttk.Entry(p_input, width=8)
        name_ent.grid(row=0, column=1, padx=5)
        ttk.Label(p_input, text="자격:").grid(row=0, column=2, padx=5)
        qual_ent = ttk.Entry(p_input, width=15)
        qual_ent.grid(row=0, column=3, padx=5)
        ttk.Label(p_input, text="취득일:").grid(row=0, column=4, padx=5)
        date_ent = ttk.Entry(p_input, width=20)
        date_ent.grid(row=0, column=5, padx=5)
        
        def add_p():
            n = name_ent.get().strip()
            if not n: return
            found = False
            for p in self.db["personnel"]:
                if p["name"] == n:
                    p["qualifications"] = qual_ent.get().strip()
                    p["date"] = date_ent.get().strip()
                    found = True; break
            if not found: self.db["personnel"].append({"name": n, "qualifications": qual_ent.get().strip(), "date": date_ent.get().strip()})
            self.save_db(); refresh_tree()
            self.personnel_names = [p["name"] for p in self.db["personnel"]]
            for cb in self.personnel_comboboxes: cb['values'] = [""] + self.personnel_names
        
        def del_p():
            selected = tree.selection()
            if not selected: return
            name_to_del = tree.item(selected[0])['values'][0]
            if messagebox.askyesno("삭제", f"{name_to_del} 님을 삭제하시겠습니까?", parent=top):
                self.db["personnel"] = [p for p in self.db["personnel"] if p["name"] != name_to_del]
                self.save_db(); refresh_tree()
                self.personnel_names = [p["name"] for p in self.db["personnel"]]
                for cb in self.personnel_comboboxes: cb['values'] = [""] + self.personnel_names
                
        ttk.Button(p_input, text="저장", command=add_p).grid(row=0, column=6, padx=5)
        ttk.Button(p_input, text="삭제", command=del_p).grid(row=0, column=7, padx=5)
        
        def on_p_sel(e):
            sel = tree.selection()
            if sel:
                item = tree.item(sel[0])['values']
                name_ent.delete(0, tk.END); name_ent.insert(0, item[0])
                qual_ent.delete(0, tk.END); qual_ent.insert(0, item[1])
                date_ent.delete(0, tk.END); date_ent.insert(0, item[2])
        tree.bind('<<TreeviewSelect>>', on_p_sel)
        
        # --- Equipment Tab ---
        e_tab = ttk.Frame(notebook)
        notebook.add(e_tab, text="장비 관리")
        
        e_cols = ("cat", "name", "spec", "qty")
        e_tree = ttk.Treeview(e_tab, columns=e_cols, show="headings", height=15)
        e_tree.heading("cat", text="구분")
        e_tree.heading("name", text="품명")
        e_tree.heading("spec", text="규격")
        e_tree.heading("qty", text="기본수량")
        e_tree.column("cat", width=80)
        e_tree.column("name", width=150)
        e_tree.column("spec", width=200)
        e_tree.column("qty", width=60)
        e_tree.pack(fill='both', expand=True, padx=10, pady=10)
        
        def refresh_e_tree():
            for item in e_tree.get_children(): e_tree.delete(item)
            for eq in self.db.get("equipment", []):
                e_tree.insert("", "end", values=(eq.get("category",""), eq.get("name",""), eq.get("spec",""), eq.get("qty","")))
        refresh_e_tree()
        
        e_input = ttk.Frame(e_tab)
        e_input.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(e_input, text="품명:").grid(row=0, column=0, padx=5)
        e_name_ent = ttk.Entry(e_input, width=15)
        e_name_ent.grid(row=0, column=1, padx=5)
        ttk.Label(e_input, text="구분:").grid(row=0, column=2, padx=5)
        e_cat_ent = ttk.Entry(e_input, width=10)
        e_cat_ent.grid(row=0, column=3, padx=5)
        ttk.Label(e_input, text="규격:").grid(row=0, column=4, padx=5)
        e_spec_ent = ttk.Entry(e_input, width=15)
        e_spec_ent.grid(row=0, column=5, padx=5)
        ttk.Label(e_input, text="수량:").grid(row=0, column=6, padx=5)
        e_qty_ent = ttk.Entry(e_input, width=5)
        e_qty_ent.grid(row=0, column=7, padx=5)
        
        def add_e():
            n = e_name_ent.get().strip()
            if not n: return
            found = False
            for eq in self.db.get("equipment", []):
                if eq["name"] == n:
                    eq["category"] = e_cat_ent.get().strip()
                    eq["spec"] = e_spec_ent.get().strip()
                    eq["qty"] = e_qty_ent.get().strip()
                    found = True; break
            if not found:
                self.db.setdefault("equipment", []).append({
                    "name": n, "category": e_cat_ent.get().strip(),
                    "spec": e_spec_ent.get().strip(), "qty": e_qty_ent.get().strip()
                })
            self.save_db(); refresh_e_tree()
            messagebox.showinfo("저장 완료", "장비 정보가 저장되었습니다. (메인 화면 갱신은 프로그램 재시작 필요)", parent=top)
            
        def del_e():
            selected = e_tree.selection()
            if not selected: return
            name_to_del = e_tree.item(selected[0])['values'][1]
            if messagebox.askyesno("삭제", f"{name_to_del} 장비를 삭제하시겠습니까?", parent=top):
                self.db["equipment"] = [eq for eq in self.db.get("equipment", []) if eq["name"] != name_to_del]
                self.save_db(); refresh_e_tree()
                
        ttk.Button(e_input, text="저장", command=add_e).grid(row=0, column=8, padx=5)
        ttk.Button(e_input, text="삭제", command=del_e).grid(row=0, column=9, padx=5)
        
        def on_e_sel(e):
            sel = e_tree.selection()
            if sel:
                item = e_tree.item(sel[0])['values']
                e_cat_ent.delete(0, tk.END); e_cat_ent.insert(0, item[0])
                e_name_ent.delete(0, tk.END); e_name_ent.insert(0, item[1])
                e_spec_ent.delete(0, tk.END); e_spec_ent.insert(0, item[2])
                e_qty_ent.delete(0, tk.END); e_qty_ent.insert(0, item[3])
        e_tree.bind('<<TreeviewSelect>>', on_e_sel)

    def save_db(self):
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(self.db, f, indent=2, ensure_ascii=False)

    def get_person_info(self, name):
        for p in self.db["personnel"]:
            if p["name"] == name:
                return p
        return None

    def save_session(self):
        session_data = {
            "template": self.tmpl_var.get(),
            "personnel": [{"name": p.get(), "role": r.get()} for p, r in zip(self.personnel_vars, self.role_vars)],
            "equipment": [{
                "db": eq["db_var"].get(),
                "qty": eq["qty_var"].get(), 
                "cat": eq["cat_var"].get(), 
                "name": eq["name_var"].get(), 
                "spec": eq["spec_var"].get()
            } for eq in self.equip_vars]
        }
        session_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "last_session.json")
        try:
            with open(session_file, "w", encoding="utf-8") as f:
                json.dump(session_data, f, ensure_ascii=False, indent=2)
        except:
            pass
            
    def load_session(self):
        session_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "last_session.json")
        if os.path.exists(session_file):
            try:
                with open(session_file, "r", encoding="utf-8") as f:
                    session_data = json.load(f)
                
                if "template" in session_data:
                    self.tmpl_var.set(session_data["template"])
                
                for i, p in enumerate(session_data.get("personnel", [])):
                    if i < len(self.personnel_vars):
                        self.personnel_vars[i].set(p.get("name", ""))
                        self.role_vars[i].set(p.get("role", ""))
                
                for i, eq in enumerate(session_data.get("equipment", [])):
                    if i < len(self.equip_vars):
                        q_val = str(eq.get("qty", ""))
                        if q_val == "0": q_val = ""
                        self.equip_vars[i]["qty_var"].set(q_val)
                        self.equip_vars[i]["cat_var"].set(eq.get("cat", ""))
                        self.equip_vars[i]["name_var"].set(eq.get("name", ""))
                        self.equip_vars[i]["spec_var"].set(eq.get("spec", ""))
                        
                        db_val = eq.get("db", "")
                        if not db_val:
                            inferred = f"[{eq.get('cat', '')}] {eq.get('name', '')}"
                            if inferred in getattr(self, 'eq_map', {}):
                                db_val = inferred
                        self.equip_vars[i]["db_var"].set(db_val)
            except:
                pass

    def generate_excel(self):
        self.save_session()
        selected_tmpl = self.tmpl_var.get()
        if not selected_tmpl:
            messagebox.showerror("오류", "템플릿을 선택하세요.")
            return
            
        tmpl_path = os.path.join(TEMPLATE_DIR, selected_tmpl)
        
        # 파일 저장 위치 묻기
        default_name = f"생성_{selected_tmpl}"
        out_path = filedialog.asksaveasfilename(
            title="엑셀 파일 저장 위치 선택",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        
        if not out_path: # 사용자가 저장을 취소한 경우
            return
            
        try:
            wb = openpyxl.load_workbook(tmpl_path)
            
            # Prepare selected personnel data
            personnel_data = []
            for i in range(15):
                name = self.personnel_vars[i].get()
                role = self.role_vars[i].get()
                if name:
                    p_info = self.get_person_info(name)
                    if p_info:
                        exp = calculate_experience(p_info.get("date", ""))
                        personnel_data.append({
                            "name": p_info["name"],
                            "role": role,
                            "exp": exp,
                            "qual": p_info.get("qualifications", ""),
                            "date": p_info.get("date", "")
                        })

            # Prepare selected equipment data
            equip_data = []
            for ev in self.equip_vars:
                try:
                    q = int(ev["qty_var"].get())
                    name = ev["name_var"].get().strip()
                    if q > 0 and name:
                        equip_data.append({
                            "category": ev["cat_var"].get(),
                            "name": name,
                            "spec": ev["spec_var"].get(),
                            "qty": q
                        })
                except ValueError:
                    pass
                    
            # 장비조직도 시트 생성 (조직도 시트 복사)
            org_sheet_name = next((s for s in wb.sheetnames if "조직도" in s), None)
            if org_sheet_name:
                eq_org_ws = wb.copy_worksheet(wb[org_sheet_name])
                eq_org_ws.title = "장비조직도"

            # Fill personnel list sheets
            for sheet_name in wb.sheetnames:
                if "인력투입계획서" in sheet_name or "인력투입변경" in sheet_name:
                    ws = wb[sheet_name]
                    # Find starting row
                    start_row = -1
                    name_col = -1
                    for row in ws.iter_rows(min_row=1, max_row=20):
                        for cell in row:
                            if cell.value == "성명":
                                start_row = cell.row + 1
                                name_col = cell.column
                                break
                        if start_row != -1: break
                    
                    if start_row != -1:
                        # 자격사항과 취득년월일 열 위치를 먼저 찾습니다. (양식마다 다름)
                        qual_col = -1
                        date_col = -1
                        for cell in ws[start_row-1]:
                            if cell.value and "자격" in str(cell.value): qual_col = cell.column
                            if cell.value and ("취득" in str(cell.value) or "년월일" in str(cell.value)): date_col = cell.column
                            
                        # 표의 마지막 줄(table_end) 동적 탐색
                        table_end = start_row
                        for r in range(start_row, start_row + 50):
                            cell = ws.cell(row=r, column=name_col)
                            b = cell.border
                            if not b or (b.left.style is None and b.right.style is None and b.bottom.style is None):
                                table_end = r - 1
                                break
                        else:
                            table_end = start_row + 30

                        # 1. 먼저 기존 데이터 지우기
                        for r in range(start_row, table_end + 1):
                            # 표 바깥 셀을 건드려 우측 페이지가 늘어나지 않도록 딱 표 영역(1~10열) 정도만 지우기
                            end_c = date_col if date_col != -1 else (qual_col if qual_col != -1 else name_col+2)
                            for c in range(1, end_c + 3):
                                cell = ws.cell(row=r, column=c)
                                if type(cell).__name__ != 'MergedCell':
                                    cell.value = ""
                            
                        # 2. 선택된 인원 데이터 중복 제거 및 담당업무 병합 후 새로 씁니다.
                        unique_personnel = []
                        seen_names = {}
                        for p in personnel_data:
                            if p["name"] in seen_names:
                                existing_p = seen_names[p["name"]]
                                if p["role"] not in existing_p["role"]:
                                    existing_p["role"] += f"\n{p['role']}"
                            else:
                                new_p = p.copy()
                                seen_names[p["name"]] = new_p
                                unique_personnel.append(new_p)

                        for idx, p in enumerate(unique_personnel):
                            row = start_row + idx
                            ws.cell(row=row, column=name_col-1).value = idx + 1 # 순번
                            ws.cell(row=row, column=name_col).value = p["name"]
                            ws.cell(row=row, column=name_col+1).value = p["role"].replace(" ", "\xa0").replace("-", "\u2011")
                            ws.cell(row=row, column=name_col+2).value = p["exp"]
                            
                            wrap_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                            
                            def expand_quals(q_str):
                                if not q_str or q_str == "-": return q_str
                                import re
                                def split_outer(s, sep):
                                    res, cur, depth = [], [], 0
                                    for c in s:
                                        if c == '(': depth += 1
                                        elif c == ')': depth -= 1
                                        elif c == sep and depth == 0:
                                            res.append("".join(cur).strip())
                                            cur = []
                                            continue
                                        cur.append(c)
                                    if cur: res.append("".join(cur).strip())
                                    return [r for r in res if r]
                                
                                expanded = []
                                for part in split_outer(q_str, '/'):
                                    for sp in split_outer(part, ','):
                                        m = re.search(r'([^\s]+)\(([^)]+)\)', sp)
                                        if m and ',' in m.group(2):
                                            prefix, inside = m.group(1), m.group(2)
                                            for sub in inside.split(','):
                                                expanded.append(sp[:m.start()] + f"{prefix}({sub.strip()})" + sp[m.end():].strip())
                                        else:
                                            expanded.append(sp.strip())
                                return "\n".join(expanded).replace(" ", "\xa0").replace("-", "\u2011")
                                
                            # 자격사항은 기사(RT, UT)를 기사(RT) \n 기사(UT)로 자동 분리 확장하여 깔끔하게 줄바꿈
                            qual_text = expand_quals(p["qual"])
                            
                            def expand_dates(d_str):
                                if not d_str or d_str == "-": return d_str
                                import re
                                # 오타 방어: ) 뒤에 숫자가 바로 오면 공백 자동 추가 (예: 2025(PT)2025 -> 2025(PT) 2025)
                                d_str = re.sub(r'\)(?=\d)', ') ', d_str)
                                def split_outer(s):
                                    res, cur, depth = [], [], 0
                                    for c in s:
                                        if c == '(': depth += 1
                                        elif c == ')': depth -= 1
                                        elif (c == ' ' or c == ',') and depth == 0:
                                            if cur: res.append("".join(cur).strip())
                                            cur = []
                                            continue
                                        cur.append(c)
                                    if cur: res.append("".join(cur).strip())
                                    return [r for r in res if r]
                                
                                expanded = []
                                for sp in split_outer(d_str):
                                    m = re.search(r'([^\s]+)\(([^)]+)\)', sp)
                                    if m and ',' in m.group(2):
                                        prefix, inside = m.group(1), m.group(2)
                                        for sub in inside.split(','):
                                            expanded.append(sp[:m.start()] + f"{prefix}({sub.strip()})" + sp[m.end():].strip())
                                    else:
                                        expanded.append(sp.strip())
                                return "\n".join(expanded).replace(" ", "\xa0").replace("-", "\u2011")
                                
                            # 취득년월일도 2025.09.26(PT,MT) 패턴을 2025.09.26(PT) \n 2025.09.26(MT)로 자동 분리 확장
                            date_text = expand_dates(p["date"])
                            # (기존의 들쭉날쭉한 행 높이 자동 맞춤 기능 제거)
                            
                            if qual_col != -1:
                                cell = ws.cell(row=row, column=qual_col)
                                lines = qual_text.split('\n') if qual_text else []
                                if any(len(line) > 18 for line in lines):
                                    small_font = InlineFont(rFont="맑은 고딕", sz=8.5)
                                    normal_font = InlineFont(rFont="맑은 고딕", sz=11)
                                    rt_elements = []
                                    for idx, line in enumerate(lines):
                                        suffix = "\n" if idx < len(lines) - 1 else ""
                                        if len(line) > 18:
                                            rt_elements.append(TextBlock(small_font, line + suffix))
                                        else:
                                            rt_elements.append(TextBlock(normal_font, line + suffix))
                                    cell.value = CellRichText(rt_elements)
                                else:
                                    cell.value = qual_text
                                cell.alignment = wrap_alignment
                                
                            if date_col != -1:
                                cell = ws.cell(row=row, column=date_col)
                                cell.value = date_text
                                cell.alignment = wrap_alignment
                                
                        # 2.5 셀 높이 일정하게 맞추기 & 남는 빈 줄 숨기기
                        max_lines = 1
                        end_c = date_col if date_col != -1 else (qual_col if qual_col != -1 else name_col+2)
                        for r in range(start_row, start_row + len(unique_personnel)):
                            for c in range(1, end_c + 3):
                                val = ws.cell(row=r, column=c).value
                                if val and isinstance(val, str):
                                    max_lines = max(max_lines, val.count('\n') + 1)
                                elif val and type(val).__name__ == 'CellRichText':
                                    text_str = "".join([tb.text for tb in val])
                                    max_lines = max(max_lines, text_str.count('\n') + 1)
                        
                        uniform_height = max(20, max_lines * 14.5)
                        
                        for r in range(start_row, table_end + 1):
                            if r < start_row + len(unique_personnel):
                                ws.row_dimensions[r].height = uniform_height
                            else:
                                ws.row_dimensions[r].hidden = True
                                
                        # 3. 인력 탭 인쇄 시 좌우 여백 틀어짐 방지 (가로 가운데 맞춤, 인쇄 영역 고정 및 좌우 여백 대칭 맞춤)
                        try:
                            from openpyxl.utils import get_column_letter
                            ws.page_margins.left = 0.5
                            ws.page_margins.right = 0.5
                            ws.print_options.horizontalCentered = True
                            
                            # 비고 열까지만 인쇄 영역 설정 (취득년월일 옆 1~2칸)
                            end_c = date_col if date_col != -1 else (qual_col if qual_col != -1 else name_col+2)
                            last_col_letter = get_column_letter(end_c + 1)
                            # 표 끝부분(table_end)까지만 인쇄되도록 설정 (하드코딩 35로 인해 빈 페이지 2페이지가 추가되는 것 방지)
                            ws.print_area = f"A1:{last_col_letter}{table_end if table_end > start_row else ws.max_row}"
                        except Exception:
                            pass
                                
                # Fill org chart sheets
                if "조직도" in sheet_name and "장비" not in sheet_name:
                    ws = wb[sheet_name]
                    
                    def set_border(r, c, left=None, right=None, top=None, bottom=None):
                        cell = ws.cell(row=r, column=c)
                        b = cell.border
                        cell.border = openpyxl.styles.Border(
                            left=left or b.left, right=right or b.right,
                            top=top or b.top, bottom=bottom or b.bottom
                        )

                    # 1. 시트 전체 초기화 (1행부터)
                    # 시트 내 기존 도형(Shape) 및 이미지 완전 제거 (그룹화된 기존 조직도 겹침 완벽 방지)
                    ws._images = []
                    
                    # 혹시나 남아있을 조건부 서식 제거 (노란색 등 엉뚱한 배경색 방지)
                    try:
                        ws.conditional_formatting = openpyxl.formatting.formatting.ConditionalFormattingList()
                    except:
                        pass
                    
                    # 시트 내 모든 병합된 셀 해제
                    merges = list(ws.merged_cells.ranges)
                    for merge in merges:
                        ws.unmerge_cells(range_string=str(merge))
                    
                    # 1행부터 값, 테두리, 배경색 지우기 (불필요하게 50열까지 넓혀서 우측으로 페이지가 늘어나지 않도록 max_column 까지만)
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            cell.value = None
                            cell.style = 'Normal'  # 강력한 스타일 초기화 (기존 노란색 등 완벽 제거)
                            cell.border = openpyxl.styles.Border()
                            cell.fill = openpyxl.styles.PatternFill(fill_type=None)

                    # 2. 인원 데이터 모델링
                    manager = []
                    safety = []
                    teams = {}
                    assistants = []
                    
                    for p in personnel_data:
                        role = p["role"]
                        if "대리인" in role or "책임" in role:
                            manager.append(p)
                        elif "안전" in role or "품질" in role:
                            safety.append(p)
                        else:
                            if "팀장" in role:
                                team_name = role.replace("팀장", "").strip()
                                if team_name not in teams: teams[team_name] = {"leader": [], "inspectors": [], "assistants": []}
                                teams[team_name]["leader"].append(p)
                            elif "검사자" in role or "검사원" in role:
                                team_name = role.replace("검사자", "").replace("검사원", "").strip()
                                if team_name not in teams: teams[team_name] = {"leader": [], "inspectors": [], "assistants": []}
                                teams[team_name]["inspectors"].append(p)
                            elif "보조" in role:
                                team_name = role.replace("검사보조", "").replace("보조", "").strip()
                                if not team_name: team_name = "PAUT"
                                if team_name not in teams: teams[team_name] = {"leader": [], "inspectors": [], "assistants": []}
                                teams[team_name]["assistants"].append(p)
                    
                    # 3. 박스 및 선 그리기 내부 함수
                    def draw_box(start_row, start_col, role_name, names, width=4):
                        if not names: names = ["-"]
                        
                        # 직책 상단 헤더
                        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+width-1)
                        header = ws.cell(row=start_row, column=start_col)
                        header.value = role_name
                        header.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                        header.font = Font(bold=True)
                        header.fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        
                        for c in range(start_col, start_col+width):
                            set_border(start_row, c, 
                                top=openpyxl.styles.Side(style='medium'), bottom=openpyxl.styles.Side(style='thin'),
                                left=openpyxl.styles.Side(style='medium') if c == start_col else None,
                                right=openpyxl.styles.Side(style='medium') if c == start_col+width-1 else None
                            )
                        
                        curr_row = start_row + 1
                        for idx, name in enumerate(names):
                            ws.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row, end_column=start_col+width-1)
                            n_cell = ws.cell(row=curr_row, column=start_col)
                            n_cell.value = name
                            n_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                            
                            is_last = (idx == len(names) - 1)
                            for c in range(start_col, start_col+width):
                                set_border(curr_row, c,
                                    left=openpyxl.styles.Side(style='medium') if c == start_col else None,
                                    right=openpyxl.styles.Side(style='medium') if c == start_col+width-1 else None,
                                    bottom=openpyxl.styles.Side(style='medium') if is_last else None
                                )
                            curr_row += 1
                        return curr_row - 1
                    
                    def draw_vline(start_row, end_row, col):
                        for r in range(start_row, end_row + 1):
                            set_border(r, col, left=openpyxl.styles.Side(style='medium'))
                            
                    def draw_hline(row, start_col, end_col):
                        for c in range(start_col, end_col + 1):
                            set_border(row, c, top=openpyxl.styles.Side(style='medium'))

                    # 4. 조직도 동적 구성 및 배치
                    BLOCK_WIDTH = 6
                    SPACING = 2
                    
                    blocks = []
                    # 방사선안전관리자는 현장대리인과 하위 조원 사이에 수직으로 단독 배치되므로 blocks에서 제외
                    for t_name, t_data in teams.items():
                        blocks.append(("팀", (t_name, t_data)))
                    # 검사보조는 이제 팀 내부로 통합되었으므로 별도 블록(보조)으로 추가하지 않음
                    num_blocks = max(1, len(blocks))
                    TOTAL_CHART_WIDTH = num_blocks * BLOCK_WIDTH + (num_blocks - 1) * SPACING
                    
                    if safety:
                        # 방사선안전관리자가 우측으로 빠질 공간 확보를 위해 필요시 너비 확장
                        req_half_width = BLOCK_WIDTH // 2 + SPACING + BLOCK_WIDTH
                        if TOTAL_CHART_WIDTH < req_half_width * 2:
                            TOTAL_CHART_WIDTH = req_half_width * 2
                            
                    target_max_col = max(18, TOTAL_CHART_WIDTH + 2)
                    
                    # 완벽한 중앙 정렬을 위한 여백 계산
                    left_margin = (target_max_col - TOTAL_CHART_WIDTH) // 2
                    
                    start_cols = [1 + left_margin + i * (BLOCK_WIDTH + SPACING) for i in range(num_blocks)]
                    
                    manager_col = 1 + left_margin + TOTAL_CHART_WIDTH // 2 - (BLOCK_WIDTH // 2)
                    if manager_col < 1: manager_col = 1
                    
                    # 모든 사용 열의 너비를 강제로 고정하여 우측 팽창 방지 및 대칭 정렬 확보
                    for col_idx in range(1, target_max_col + 1):
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        ws.column_dimensions[col_letter].width = 4.2
                        
                    # 제목은 1열부터 인쇄/뷰 영역의 끝(target_max_col)까지 병합하여 완벽하게 중앙에 위치시킴
                    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=target_max_col)
                    title_cell = ws.cell(row=2, column=1)
                    title_cell.value = "현 장 조 직 도"
                    title_cell.font = Font(size=24, bold=True)
                    title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if not manager and not blocks:
                        pass # 그릴 인원이 없음
                    else:
                        # (1) 현장대리인
                        manager_names = [p["name"] for p in manager] if manager else ["-"]
                        m_bottom = draw_box(5, manager_col, "현장대리인", manager_names, width=BLOCK_WIDTH)
                        m_center = manager_col + BLOCK_WIDTH // 2
                        
                        # (1-1) 방사선안전관리자 (우측 빼기)
                        if safety:
                            # 현장대리인에서 가로선이 나가는 곳까지 수직선
                            draw_vline(m_bottom + 1, m_bottom + 1, m_center)
                            
                            # 우측으로 수평선 긋기
                            safety_col = manager_col + BLOCK_WIDTH + SPACING
                            safety_center = safety_col + BLOCK_WIDTH // 2
                            draw_hline(m_bottom + 2, m_center, safety_center - 1)
                            
                            # 우측 끝에서 아래로 수직선
                            draw_vline(m_bottom + 2, m_bottom + 2, safety_center)
                            
                            # 방사선안전관리자 박스 그리기
                            safety_names = [p["name"] for p in safety]
                            s_bottom = draw_box(m_bottom + 3, safety_col, "방사선안전관리자", safety_names, width=BLOCK_WIDTH)
                            
                            # 현장대리인 중앙 기둥(트렁크)을 방사선안전관리자 박스 아래까지 쭉 내리기
                            draw_vline(m_bottom + 2, s_bottom, m_center)
                            
                            m_bottom = s_bottom
                            
                        last_row_approx = m_bottom + 15
                        ws.print_area = f"A1:{openpyxl.utils.get_column_letter(target_max_col)}{last_row_approx}"
                        
                        # 인쇄 시 1페이지 너비에 딱 맞게 자동 축소되도록 강제 설정 (우측 점선 발생 방지)
                        ws.sheet_properties.pageSetUpPr.fitToPage = True
                        ws.page_setup.fitToWidth = 1
                        ws.page_setup.fitToHeight = 0
                        
                        if blocks:
                            # (2) 연결 선 (가로)
                            first_center = start_cols[0] + BLOCK_WIDTH // 2
                            last_center = start_cols[-1] + BLOCK_WIDTH // 2
                            if num_blocks > 1:
                                # 가로선이 오른쪽 끝 수직선 밖으로 삐져나가지 않도록 last_center - 1 까지만 그립니다.
                                draw_hline(m_bottom + 2, first_center, last_center - 1)
                            
                            # 현장대리인에서 가로선으로 내려오는 세로선 (가로선을 관통하지 않도록 m_bottom+1까지만)
                            draw_vline(m_bottom + 1, m_bottom + 1, m_center)
                            
                            # (3) 하위 부서들
                            for i, block in enumerate(blocks):
                                b_type, b_data = block
                                b_col = start_cols[i]
                                b_center = b_col + BLOCK_WIDTH // 2
                                
                                # 가로선에서 각 부서로 내려가는 세로선
                                draw_vline(m_bottom + 2, m_bottom + 3, b_center)
                                
                                start_r = m_bottom + 4
                                if b_type == "팀":
                                    t_name, t_data = b_data
                                    leaders = [p["name"] for p in t_data.get("leader", [])]
                                    inspectors = [p["name"] for p in t_data.get("inspectors", [])]
                                    team_assistants = [p["name"] for p in t_data.get("assistants", [])]
                                    
                                    b_bottom = draw_box(start_r, b_col, f"{t_name} 팀장", leaders, width=BLOCK_WIDTH)
                                    if inspectors:
                                        draw_vline(b_bottom + 1, b_bottom + 2, b_center)
                                        b_bottom = draw_box(b_bottom + 3, b_col, f"{t_name} 검사자", inspectors, width=BLOCK_WIDTH)
                                    if team_assistants:
                                        draw_vline(b_bottom + 1, b_bottom + 2, b_center)
                                        b_bottom = draw_box(b_bottom + 3, b_col, f"{t_name} 검사보조", team_assistants, width=BLOCK_WIDTH)
                                        
                        # 조직도 그리기가 모두 끝난 후, 우측으로 빈 공간이 무한정 늘어나는 것을 방지하기 위해 잉여 열 완전 삭제
                        if ws.max_column > target_max_col:
                            ws.delete_cols(target_max_col + 1, ws.max_column - target_max_col + 5)
                            
                        # 인쇄 영역을 실제로 그려진 범위(ws.max_row)에 딱 맞게 재설정
                        ws.print_area = f"A1:{openpyxl.utils.get_column_letter(target_max_col)}{ws.max_row + 2}"
                        if ws.sheet_properties.pageSetUpPr is None:
                            ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties()
                        ws.sheet_properties.pageSetUpPr.fitToPage = True
                        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                        ws.page_setup.fitToWidth = 1
                        ws.page_setup.fitToHeight = 1
                        ws.print_options.horizontalCentered = True
                        ws.print_options.verticalCentered = True
                                        
                # Fill eq org chart
                elif "장비조직도" in sheet_name:
                    ws = wb[sheet_name]
                    
                    def set_border(r, c, left=None, right=None, top=None, bottom=None):
                        cell = ws.cell(row=r, column=c)
                        b = cell.border
                        cell.border = openpyxl.styles.Border(
                            left=left or b.left, right=right or b.right,
                            top=top or b.top, bottom=bottom or b.bottom
                        )
                    
                    ws._images = []
                    try: ws.conditional_formatting = openpyxl.formatting.formatting.ConditionalFormattingList()
                    except: pass
                    merges = list(ws.merged_cells.ranges)
                    for merge in merges: ws.unmerge_cells(range_string=str(merge))
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            cell.value = None
                            cell.style = 'Normal'
                            cell.border = openpyxl.styles.Border()
                            cell.fill = openpyxl.styles.PatternFill(fill_type=None)
                            
                    eq_groups = {}
                    for eq in equip_data:
                        cat = eq["category"].strip()
                        if not cat: cat = "기타"
                        eq_groups.setdefault(cat, []).append(eq)
                        
                    def draw_eq_box(start_row, start_col, title, items, width=4):
                        if not items: return start_row
                        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+width-1)
                        header = ws.cell(row=start_row, column=start_col)
                        header.value = title
                        header.alignment = Alignment(horizontal='center', vertical='center')
                        header.font = Font(bold=True)
                        header.fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        for c in range(start_col, start_col+width):
                            set_border(start_row, c, 
                                top=openpyxl.styles.Side(style='medium'), bottom=openpyxl.styles.Side(style='medium'),
                                left=openpyxl.styles.Side(style='medium') if c == start_col else None,
                                right=openpyxl.styles.Side(style='medium') if c == start_col+width-1 else None
                            )
                        
                        curr_row = start_row + 1
                        for idx, eq in enumerate(items):
                            ws.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row+2, end_column=start_col+width-1)
                            n_cell = ws.cell(row=curr_row, column=start_col)
                            
                            rt_elements = [
                                TextBlock(InlineFont(rFont='맑은 고딕', b=True, sz=11), f"{eq['name']}\n"),
                                TextBlock(InlineFont(rFont='맑은 고딕', sz=9, color=openpyxl.styles.colors.Color(rgb='FF595959')), f"({eq['spec']})\n"),
                                TextBlock(InlineFont(rFont='맑은 고딕', b=True, color=openpyxl.styles.colors.Color(rgb='FF0000FF')), f"{eq['qty']}개")
                            ]
                            n_cell.value = CellRichText(rt_elements)
                            n_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            
                            is_last = (idx == len(items) - 1)
                            for r_idx in range(3):
                                r = curr_row + r_idx
                                for c in range(start_col, start_col+width):
                                    set_border(r, c,
                                        left=openpyxl.styles.Side(style='medium') if c == start_col else None,
                                        right=openpyxl.styles.Side(style='medium') if c == start_col+width-1 else None,
                                        bottom=openpyxl.styles.Side(style='medium') if (is_last and r_idx == 2) else openpyxl.styles.Side(style='dotted') if r_idx == 2 else None
                                    )
                            curr_row += 3
                        return curr_row - 1
                        
                    def draw_vline(r1, r2, col):
                        for r in range(r1, r2 + 1):
                            set_border(r, col, left=openpyxl.styles.Side(style='medium'))
                            
                    def draw_hline(row, c1, c2):
                        for c in range(min(c1, c2), max(c1, c2)):
                            set_border(row, c, top=openpyxl.styles.Side(style='medium'))
                            
                    BLOCK_WIDTH = 6
                    SPACING = 2
                    
                    cats = list(eq_groups.keys())
                    if cats:
                        # Calculate centers first to perfect align the root box
                        curr_col = 2
                        centers = []
                        for _ in cats:
                            centers.append(curr_col + (BLOCK_WIDTH // 2))
                            curr_col += BLOCK_WIDTH + SPACING
                            
                        if len(centers) > 1:
                            mid_boundary = (min(centers) + max(centers)) // 2
                        else:
                            mid_boundary = centers[0]
                            
                        root_col = max(2, mid_boundary - (BLOCK_WIDTH // 2))
                        root_center = root_col + (BLOCK_WIDTH // 2)
                        root_row = 2
                        
                        ws.merge_cells(start_row=root_row, start_column=root_col, end_row=root_row+1, end_column=root_col+BLOCK_WIDTH-1)
                        title_cell = ws.cell(row=root_row, column=root_col)
                        title_cell.value = "장비 투입 현황"
                        title_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                        title_cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
                        for r in range(root_row, root_row+2):
                            for c in range(root_col, root_col+BLOCK_WIDTH):
                                set_border(r, c, 
                                    left=openpyxl.styles.Side(style='medium') if c == root_col else None,
                                    right=openpyxl.styles.Side(style='medium') if c == root_col+BLOCK_WIDTH-1 else None,
                                    top=openpyxl.styles.Side(style='medium') if r == root_row else None,
                                    bottom=openpyxl.styles.Side(style='medium') if r == root_row+1 else None
                                )
                        
                        branch_row = root_row + 4
                        max_bottom = branch_row
                        
                        temp_col = 2
                        for cat in cats:
                            b = draw_eq_box(branch_row, temp_col, f"[{cat}] 장비", eq_groups[cat], width=BLOCK_WIDTH)
                            if b > max_bottom: max_bottom = b
                            temp_col += BLOCK_WIDTH + SPACING
                            
                        draw_vline(root_row + 2, branch_row - 2, root_center)
                        
                        if len(centers) > 1:
                            draw_hline(branch_row - 1, min(centers), max(centers))
                        for c in centers:
                            draw_vline(branch_row - 1, branch_row - 1, c)
                                
                        target_max_col = temp_col - SPACING - 1
                        if target_max_col < 10: target_max_col = 10
                        if ws.max_column > target_max_col:
                            ws.delete_cols(target_max_col + 1, ws.max_column - target_max_col + 5)
                            
                        # Set uniform column widths and fit to 1 page landscape
                        if ws.sheet_properties.pageSetUpPr is None:
                            ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties()
                        ws.sheet_properties.pageSetUpPr.fitToPage = True
                        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                        ws.page_setup.fitToWidth = 1
                        ws.page_setup.fitToHeight = 1
                        for i in range(1, target_max_col + 1):
                            col_letter = openpyxl.utils.get_column_letter(i)
                            ws.column_dimensions[col_letter].width = 3.5
                            
                        ws.print_area = f"A1:{openpyxl.utils.get_column_letter(target_max_col)}{max_bottom + 1}"

                # Fill equipment sheets
                elif "장비" in sheet_name and "조직도" not in sheet_name:
                    ws = wb[sheet_name]
                    
                    # Gather selected equipment from UI
                    equip_data = []
                    for ev in self.equip_vars:
                        try:
                            q = int(ev["qty_var"].get())
                            if q > 0:
                                equip_data.append({
                                    "category": ev["cat_var"].get(),
                                    "name": ev["name_var"].get(),
                                    "spec": ev["spec_var"].get(),
                                    "qty": q
                                })
                        except ValueError:
                            pass
                            
                    start_row = -1
                    name_col = -1
                    
                    # Find header row containing "품명", "장비명", or "규격"
                    for row in ws.iter_rows(min_row=1, max_row=20):
                        for cell in row:
                            if cell.value:
                                val = str(cell.value).replace(" ", "").replace("\n", "")
                                if "품명" in val or "장비명" in val:
                                    start_row = cell.row + 1
                                    name_col = cell.column
                                    break
                        if start_row != -1: break
                        
                    if start_row != -1 and name_col != -1:
                        cat_col = name_col - 1
                        spec_col = name_col + 1
                        qty_col = name_col + 2
                        time_col = name_col + 3
                        
                        # 0. 표의 마지막 줄(table_end) 동적 탐색 (결재란 등 푸터 영역 침범 방지)
                        table_end = start_row
                        for r in range(start_row, start_row + 50):
                            cell = ws.cell(row=r, column=name_col)
                            b = cell.border
                            # 표 내부 셀은 보통 테두리가 있음. 좌/우/아래 테두리가 모두 없으면 표가 끝났다고 판단.
                            if not b or (b.left.style is None and b.right.style is None and b.bottom.style is None):
                                table_end = r - 1
                                break
                        else:
                            table_end = start_row + 30
                            
                        # 1. 표 내부의 병합만 완벽 해제 (푸터 영역은 건드리지 않음)
                        merged_ranges = list(ws.merged_cells.ranges)
                        for m_range in merged_ranges:
                            if m_range.min_row >= start_row and m_range.max_row <= table_end:
                                ws.unmerge_cells(str(m_range))
                                    
                        # 2. 표 내부 기존 데이터 지우기 및 남는 줄 숨김 처리
                        for r in range(start_row, table_end + 1):
                            # 표 바깥 셀을 건드리면 우측으로 빈 페이지(2, 3페이지)가 무한 증식하므로 딱 표 영역(time_col+1)까지만 지우기
                            for c in range(cat_col, time_col + 2):  
                                cell = ws.cell(row=r, column=c)
                                if type(cell).__name__ != 'MergedCell':
                                    cell.value = ""
                                cell.border = openpyxl.styles.Border()
                            # 입력된 장비 개수보다 아래에 있는 표 빈 줄은 엑셀에서 아예 숨김 처리 (1페이지 핏)
                            if r >= start_row + len(equip_data):
                                ws.row_dimensions[r].hidden = True
                            
                        # 3. Insert new data
                        thin_border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'), 
                            right=openpyxl.styles.Side(style='thin'), 
                            top=openpyxl.styles.Side(style='thin'), 
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        for idx, eq in enumerate(equip_data):
                            row = start_row + idx
                            ws.cell(row=row, column=cat_col).value = eq["category"]
                            ws.cell(row=row, column=name_col).value = eq["name"]
                            # 규격 텍스트가 애매하게 자동 줄바꿈되는 것을 막고, 쉼표나 슬래시를 기준으로 깔끔하게 두 줄로 나눔 (공백 유무 상관없이 완벽하게 분리)
                            spec_text = eq["spec"]
                            if spec_text:
                                import re
                                spec_text = re.sub(r'\s*[,/]\s*', '\n', spec_text)
                            ws.cell(row=row, column=spec_col).value = spec_text
                            ws.cell(row=row, column=qty_col).value = str(eq["qty"])
                            
                            # 첫 줄은 전체 텍스트, 두 번째 줄부터는 '"' (상동) 표시
                            if idx == 0:
                                ws.cell(row=row, column=time_col).value = "공정상황에\n따라 적시투입"
                                # '입' 글자가 잘리지 않도록 투입시기 열의 너비를 살짝만 (14) 넓힘
                                try:
                                    from openpyxl.utils import get_column_letter
                                    ws.column_dimensions[get_column_letter(time_col)].width = 14
                                except:
                                    pass
                            else:
                                ws.cell(row=row, column=time_col).value = '"'
                            
                            wrap_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                            
                            # 여백 삐져나감 방지: F열(비고)까지만 딱 맞게 테두리 생성 (time_col+2)
                            for c in range(cat_col, time_col+2):
                                try:
                                    ws.cell(row=row, column=c).alignment = wrap_alignment
                                    ws.cell(row=row, column=c).border = thin_border
                                except Exception: pass
                                
                        # 3. 새 데이터에 맞춰 카테고리 열(구분) 다시 병합
                        if equip_data:
                            start_merge = start_row
                            current_cat = equip_data[0]["category"]
                            for r in range(start_row + 1, start_row + len(equip_data) + 1):
                                cat = equip_data[r - start_row]["category"] if r < start_row + len(equip_data) else None
                                if cat != current_cat:
                                    if r - 1 > start_merge:
                                        ws.merge_cells(start_row=start_merge, start_column=cat_col, end_row=r-1, end_column=cat_col)
                                    start_merge = r
                                    current_cat = cat
                                    
                        # 4. 인쇄 시 좌우 여백 틀어짐 방지 (가로 가운데 맞춤, 인쇄 영역 고정 및 좌우 여백 대칭 맞춤)
                        try:
                            from openpyxl.utils import get_column_letter
                            ws.page_margins.left = 0.5
                            ws.page_margins.right = 0.5
                            ws.print_options.horizontalCentered = True
                            last_col_letter = get_column_letter(time_col + 1)
                            ws.print_area = f"A1:{last_col_letter}{table_end if table_end > start_row else ws.max_row}"
                        except Exception:
                            pass
                                        
            # 3개 카테고리(인력, 장비, 조직도)별로 딱 1개의 시트만 남기고 이름 깔끔하게 수정
            keep_sheets = {}
            for sheet_name in wb.sheetnames:
                if "인력" in sheet_name and "인력" not in keep_sheets:
                    keep_sheets["인력"] = sheet_name
                elif "장비조직도" in sheet_name and "장비조직도" not in keep_sheets:
                    keep_sheets["장비조직도"] = sheet_name
                elif "장비" in sheet_name and "조직도" not in sheet_name and "장비" not in keep_sheets:
                    keep_sheets["장비"] = sheet_name
                elif "조직도" in sheet_name and "장비" not in sheet_name and "조직도" not in keep_sheets:
                    keep_sheets["조직도"] = sheet_name
                    
            # 채택되지 않은 나머지 찌꺼기/중복 시트 모두 삭제
            for sheet_name in wb.sheetnames:
                if sheet_name not in keep_sheets.values():
                    del wb[sheet_name]
            
            # 남은 시트들의 이름을 숫자 '(2)' 등이 없는 깔끔한 고정 이름으로 강제 변경
            if "인력" in keep_sheets:
                wb[keep_sheets["인력"]].title = "인력투입계획서"
            if "장비" in keep_sheets:
                wb[keep_sheets["장비"]].title = "장비투입계획서"
            if "조직도" in keep_sheets:
                wb[keep_sheets["조직도"]].title = "조직도"
            if "장비조직도" in keep_sheets:
                wb[keep_sheets["장비조직도"]].title = "장비조직도"
                    
            # 모든 시트의 그룹화 상태(다중 선택)를 해제하고 첫 번째 시트만 활성화
            for i, sheet in enumerate(wb.worksheets):
                try:
                    sheet.views.sheetView[0].tabSelected = (i == 0)
                except Exception:
                    pass
            wb.active = 0
            
            wb.save(out_path)
            messagebox.showinfo("완료", f"성공적으로 생성되었습니다.\n저장 위치: {out_path}")
        except Exception as e:
            messagebox.showerror("오류", f"생성 중 오류 발생:\n{e}")

if __name__ == "__main__":
    root = tk.Tk()
    
    # Custom styles
    style = ttk.Style()
    if 'clam' in style.theme_names():
        style.theme_use('clam')
    style.configure('Accent.TButton', font=('Malgun Gothic', 10, 'bold'), background='#0052cc', foreground='white')
    
    app = DeploymentApp(root)
    
    def on_closing():
        app.save_session()
        root.destroy()
        
    root.protocol("WM_DELETE_WINDOW", on_closing)
    root.mainloop()
