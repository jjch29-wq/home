import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image
import os
import json
from datetime import datetime

CONFIG_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "config_hazard.json")

# ───────────────────────── 기본 내용 ─────────────────────────
DEFAULT_UTPT_LEFT = (
    "1. 밀폐된 공간이나 환기가 불충분한 곳에서 세척액 및 침투액 사용 시 유기용제 증기 흡입에 의한 중독 위험(최대)\n"
    "2. [PT] 인화성 에어로졸(세척액, 현상액) 취급 중 주변 용접 불꽃 등 화기 접촉으로 인한 화재 및 폭발 위험(최대)\n"
    "3. [PT] 침투액 등 화학물질 취급 시 보호구 미착용으로 인한 피부 접촉 및 피부염 발생 위험(중)\n"
    "4. [UT] 고소작업(비계, 사다리 등) 중 초음파탐상검사 시 안전대 미체결 및 부주의에 의한 추락 위험(최대)\n"
    "5. [UT] 불편한 자세로 장시간 탐촉자를 문지르는 반복작업에 의한 근골격계 질환 위험(중)\n"
    "6. [UT] 접촉매질(Couplant, 겔 등)이 바닥에 흘러 작업자가 밟고 미끄러짐(넘어짐) 위험(중)\n"
    "7. [UT] 탐상설비 전원 케이블 피복 손상 또는 습윤한 환경에서 작업 시 감전 위험(중)\n"
    "8. 통제되지 않은 타공정과 혼재작업 시 중량물 낙하 등에 의한 맞음 사고 위험(최대)\n"
    "9. 협소한 밀폐공간 내 작업 질식 또는 약품 흡입 위험(최대)\n"
    "10. 야간작업 시, 어두운 환경에서 이동 및 작업할 경우 시야 미확보에 의한 넘어짐 및 부딪힘 위험(최대)\n"
    "11. 폐기물(사용한 결재, 빈 캔)의 무단 방치 및 화기 근접으로 인한 화재 위험(중)\n"
    "12. 폭염 시 야외·옥상 등 고온 환경에서 작업할 경우 열사병·열탈진 등 온열질환 발생 위험(중)\n"
    "13. 한파 시 저온 환경에서 장시간 작업할 경우 저체온증·동상 등 한랭질환 발생 위험(중)"
)
DEFAULT_UTPT_RIGHT = (
    "1. ② 중독\n"
    "2. ⑩ 폭발·파열, ⑪ 화재\n"
    "3. ⑭ 화학물질 누출·접촉\n"
    "4. ① 떨어짐\n"
    "5. ⑫ 불균형 및 무리한 동작\n"
    "6. ② 넘어짐\n"
    "7. ⑨ 감전\n"
    "8. ⑤ 맞음\n"
    "9. ② 중독\n"
    "10. ② 넘어짐, ④ 부딪힘\n"
    "11. ⑪ 화재\n"
    "12. ⑬ 이상온도·물체접촉(온열질환)\n"
    "13. ⑬ 이상온도·물체접촉(한랭질환)"
)

DEFAULT_RT_LEFT = (
    "1. 작업현장 내 발생된 물기 및 폐기물 등에 의해 넘어짐 위험(중)\n"
    "2. 자재 및 장비 이동시 과도한 무게 취급에 따른 근로자 근골격계 질환 발생 위험(중)\n"
    "3. 가이딩 찍힘, 꺾임 등으로 방사선원의 이탈 및 회수가 되지 않는 사고 위험(최대)\n"
    "4. 방사선조사기 Pig Tail 유격으로 인한 방사선원 이탈 사고(최대)\n"
    "5. 방사선관리구역 미설정 시 일반인 미통제로 피폭사고 위험(최대)\n"
    "6. 방사선 개인안전장구 미착용으로 방사선 노출시 방사선 피폭을 인지하지 못하여 과피폭 위험(최대)\n"
    "7. 교정되지 않은 서베이메타 사용시 실제 방사선량 측정을 할 수 없을시 피폭사고 위험(최대)\n"
    "8. 콜리메타 미사용 또는 부적절한 콜리메타 사용 시 피폭사고 위험(최대)\n"
    "9. 통제되지 않은 타공정과 혼재작업 시 일반인 피폭사고 위험(최대)\n"
    "10. 무거운 방사선장비 이동 및 사용시 방사선장비의 낙하로 장비 파손으로 인한 방사선사고 위험(최대)\n"
    "11. 협소한 밀폐공간 내 작업시 질식 또는 약품 흡입 위험(최대)\n"
    "12. 야간작업 시, 어두운 환경에서 이동 및 작업할 경우 넘어짐 또는 방사선사고 발생 위험(최대)\n"
    "13. 필름현상 시, MSDS물질을 취급함에 따라 흡입 위험(중)\n"
    "14. 폭염 시 야외·옥상 등 고온 환경에서 작업할 경우 열사병·열탈진 등 온열질환 발생 위험(중)\n"
    "15. 한파 시 저온 환경에서 장시간 작업할 경우 저체온증·동상 등 한랭질환 발생 위험(중)"
)
DEFAULT_RT_RIGHT = (
    "1. ② 넘어짐\n"
    "2. ⑫ 불균형 및 무리한 동작\n"
    "3. ⑮ 기타(방사선피폭)\n"
    "4. ⑮ 기타(방사선피폭)\n"
    "5. ⑮ 기타(방사선피폭)\n"
    "6. ⑮ 기타(방사선피폭)\n"
    "7. ⑮ 기타(방사선피폭)\n"
    "8. ⑮ 기타(방사선피폭)\n"
    "9. ⑮ 기타(방사선피폭)\n"
    "10. ⑤ 맞음, ⑮ 기타(방사선피폭)\n"
    "11. ② 중독\n"
    "12. ② 넘어짐\n"
    "13. ② 중독\n"
    "14. ⑬ 이상온도·물체접촉(온열질환)\n"
    "15. ⑬ 이상온도·물체접촉(한랭질환)"
)

# ───────────────────────── config ─────────────────────────
def load_config():
    if os.path.exists(CONFIG_FILE_PATH):
        try:
            with open(CONFIG_FILE_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}

def save_config(config):
    os.makedirs(os.path.dirname(CONFIG_FILE_PATH), exist_ok=True)
    try:
        with open(CONFIG_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=4)
    except:
        pass

# ───────────────────────── Excel 스타일 ─────────────────────────
bold_font      = Font(name='맑은 고딕', size=11, bold=True)
title_font     = Font(name='맑은 고딕', size=14, bold=True)
box_title_font = Font(name='맑은 고딕', size=16, bold=True)
normal_font    = Font(name='맑은 고딕', size=11)

center_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
left_top_align = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
right_align    = Alignment(horizontal="right",  vertical="center", wrap_text=True)

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'),  bottom=Side(style='thin'))

def set_border(ws, min_col, min_row, max_col, max_row):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border

# ───────────────────────── Excel 생성 ─────────────────────────
def create_excel(output_filename, data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # 기본 시트 제거

    sheets = [
        ("UT·PT 조사표", data["utpt"]),
        ("RT 조사표",    data["rt"]),
    ]

    for sheet_title, sheet_data in sheets:
        ws = wb.create_sheet(title=sheet_title)

        # 페이지 설정
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation  = 'portrait'
        ws.page_setup.paperSize    = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage    = True
        ws.page_setup.fitToWidth   = 1
        ws.page_setup.fitToHeight  = 1
        ws.page_margins.left   = 0.2
        ws.page_margins.right  = 0.2
        ws.page_margins.top    = 0.39
        ws.page_margins.bottom = 0.39
        ws.page_margins.header = 0
        ws.page_margins.footer = 0
        ws.print_options.horizontalCentered = True

        # 열 너비 (A=라벨, B·D·F=이름, C·E·G=(인) / 내용)
        ws.column_dimensions['A'].width = 16   # 라벨
        ws.column_dimensions['B'].width = 16   # 수행자1 이름
        ws.column_dimensions['C'].width = 5    # 수행자1 (인)
        ws.column_dimensions['D'].width = 16   # 수행자2 이름
        ws.column_dimensions['E'].width = 5    # 수행자2 (인)
        ws.column_dimensions['F'].width = 22   # 수행자3 이름 / 사고유형 내용
        ws.column_dimensions['G'].width = 8    # 수행자3 (인)

        # 행 높이 계산
        A4_H   = 785
        H1, H2, H3, H4, H5 = 22, 6, 32, 22, 22
        H6, H7, H8, H9, H20 = 28, 28, 20, 22, 120
        FIXED  = H1+H2+H3+H4+H5+H6+H7+H8+H9+H20
        ROW_H  = max(A4_H - FIXED, 200) / 10

        ws.row_dimensions[1].height  = H1
        ws.row_dimensions[2].height  = H2
        ws.row_dimensions[3].height  = H3
        ws.row_dimensions[4].height  = H4
        ws.row_dimensions[5].height  = H5
        ws.row_dimensions[6].height  = H6
        ws.row_dimensions[7].height  = H7
        ws.row_dimensions[8].height  = H8
        ws.row_dimensions[9].height  = H9
        for r in range(10, 20):
            ws.row_dimensions[r].height = ROW_H
        ws.row_dimensions[20].height = H20

        # ── 내용 ──
        ws.merge_cells('A1:G1')
        ws['A1'] = "1. 순회점검에 의한 유해·위험요인 조사"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells('A3:G3')
        ws['A3'] = "사업장 순회점검에 의한 유해·위험요인 조사표"
        ws['A3'].font = box_title_font
        ws['A3'].alignment = center_align
        ws['A3'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        ws.merge_cells('A4:A5')
        ws['A4'] = "실시방법"
        ws['A4'].font = bold_font
        ws['A4'].alignment = center_align

        ws.merge_cells('B4:G5')
        ws['B4'] = "위험성평가 수행자가 정기적으로 사업장을 순회점검하고 이 조사표를 사용하여 유해·위험요인을 찾음"
        ws['B4'].font = bold_font
        ws['B4'].alignment = left_align

        # 수행자
        ws['A6'] = "수행자 성명 : "
        ws['A6'].font = bold_font
        ws['A6'].alignment = right_align

        workers   = data.get("workers", ["", "", ""])
        signs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "signs")
        name_cols = [('B','C'), ('D','E'), ('F','G')]
        for i, (ncol, icol) in enumerate(name_cols):
            name = workers[i] if i < len(workers) else ""
            ws[f'{ncol}6'] = name
            ws[f'{ncol}6'].font = normal_font
            ws[f'{ncol}6'].alignment = center_align
            ws[f'{icol}6'] = "(인)"
            ws[f'{icol}6'].font = normal_font
            ws[f'{icol}6'].alignment = center_align
            if name:
                for fname in [f"{name}_padded.png", f"{name}.png"]:
                    sp = os.path.join(signs_dir, fname)
                    if os.path.exists(sp):
                        img = Image(sp); img.width=40; img.height=28
                        ws.add_image(img, f'{icol}6'); break

        # 감독
        ws['A7'] = "감 독 성명 : "
        ws['A7'].font = bold_font
        ws['A7'].alignment = right_align

        supervisor = data.get("supervisor", "")
        ws['B7'] = supervisor
        ws['B7'].font = normal_font
        ws['B7'].alignment = center_align
        ws['C7'] = "(인)"
        ws['C7'].font = normal_font
        ws['C7'].alignment = center_align
        if supervisor:
            for fname in [f"{supervisor}_padded.png", f"{supervisor}.png"]:
                sp = os.path.join(signs_dir, fname)
                if os.path.exists(sp):
                    img = Image(sp); img.width=40; img.height=28
                    ws.add_image(img, 'C7'); break
        ws.merge_cells('D7:G7')

        # 수행 일시
        ws['A8'] = "수 행 일 시 : "
        ws['A8'].font = bold_font
        ws['A8'].alignment = right_align
        ws.merge_cells('B8:G8')
        ws['B8'] = "  " + data.get("eval_date", "")
        ws['B8'].font = normal_font
        ws['B8'].alignment = left_align

        # 표 헤더
        ws.merge_cells('A9:E9')
        ws['A9'] = "유해·위험작업"
        ws['A9'].font = bold_font
        ws['A9'].alignment = center_align
        ws.merge_cells('F9:G9')
        ws['F9'] = "사고, 질병의 유형"
        ws['F9'].font = bold_font
        ws['F9'].alignment = center_align

        # 표 내용 - 항목별로 행 하나씩 (병합 없이)
        left_items  = [l for l in sheet_data.get("left_text","").split("\n") if l.strip()]
        right_items = [l for l in sheet_data.get("right_text","").split("\n") if l.strip()]
        n_items = max(len(left_items), len(right_items), 1)

        # 행 높이: A4 전체 - 고정 행 - 범례 행을 항목 수로 나눔
        content_h = max(A4_H - FIXED, 200)
        item_h = content_h / n_items

        content_font = Font(name='맑은 고딕', size=9)
        for idx in range(n_items):
            row = 10 + idx
            ws.row_dimensions[row].height = item_h

            # 왼쪽: A-E 병합 (열만 병합, 행은 독립)
            ws.merge_cells(f'A{row}:E{row}')
            left_val = left_items[idx] if idx < len(left_items) else ""
            ws[f'A{row}'] = left_val
            ws[f'A{row}'].font = content_font
            ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # 오른쪽: F-G 병합 (열만 병합)
            ws.merge_cells(f'F{row}:G{row}')
            right_val = right_items[idx] if idx < len(right_items) else ""
            ws[f'F{row}'] = right_val
            ws[f'F{row}'].font = content_font
            ws[f'F{row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 범례 행 번호
        legend_row = 10 + n_items
        ws.row_dimensions[legend_row].height = H20

        # 범례
        ws.merge_cells(f'A{legend_row}:E{legend_row}')
        legend = ("사고의 유형\n"
                  "① 떨어짐   ⑨ 감전\n"
                  "② 넘어짐   ⑩ 폭발·파열\n"
                  "③ 깔림      ⑪ 화재\n"
                  "④ 부딪힘   ⑫ 불균형 및 무리한 동작\n"
                  "⑤ 맞음     ⑬ 이상온도·물체접촉\n"
                  "⑥ 무너짐   ⑭ 화학물질 누출·접촉\n"
                  "⑦ 끼임      ⑮ 기타\n"
                  "⑧ 절단·베임·찔림")
        ws[f'A{legend_row}'] = legend
        ws[f'A{legend_row}'].font = Font(name='맑은 고딕', size=9)
        ws[f'A{legend_row}'].alignment = left_top_align

        ws.merge_cells(f'F{legend_row}:G{legend_row}')
        ws[f'F{legend_row}'] = "질병의 유형\n① 진폐\n② 중독\n③ 난청\n④ 요통\n⑤ 기타"
        ws[f'F{legend_row}'].font = Font(name='맑은 고딕', size=9)
        ws[f'F{legend_row}'].alignment = left_top_align

        # 테두리 (행 3 ~ legend_row)
        set_border(ws, 1, 3, 7, legend_row)
        medium = Side(style='medium')
        for row in range(3, legend_row + 1):
            c1 = ws.cell(row=row, column=1)
            c1.border = Border(left=medium, right=c1.border.right, top=c1.border.top, bottom=c1.border.bottom)
            c7 = ws.cell(row=row, column=7)
            c7.border = Border(right=medium, left=c7.border.left, top=c7.border.top, bottom=c7.border.bottom)
        for col in range(1, 8):
            ct = ws.cell(row=3, column=col)
            ct.border = Border(top=medium, bottom=ct.border.bottom, left=ct.border.left, right=ct.border.right)
            cb = ws.cell(row=legend_row, column=col)
            cb.border = Border(bottom=medium, top=cb.border.top, left=cb.border.left, right=cb.border.right)
        ws['A3'].border = Border(top=medium, left=medium, right=ws['A3'].border.right, bottom=ws['A3'].border.bottom)
        ws[f'G3'].border = Border(top=medium, right=medium, left=ws['G3'].border.left, bottom=ws['G3'].border.bottom)
        ws[f'A{legend_row}'].border = Border(bottom=medium, left=medium, right=ws[f'A{legend_row}'].border.right, top=ws[f'A{legend_row}'].border.top)
        ws[f'G{legend_row}'].border = Border(bottom=medium, right=medium, left=ws[f'G{legend_row}'].border.left, top=ws[f'G{legend_row}'].border.top)

    try:
        wb.save(output_filename)
        return True, f"저장 완료: {os.path.basename(output_filename)}"
    except Exception as e:
        return False, f"저장 오류: {e}"



# ───────────────────────── GUI ─────────────────────────
class HazardApp:
    def __init__(self, root):
        self.root = root
        self.root.title("유해·위험요인 조사표 자동 생성기")
        self.root.geometry("850x850")

        style = ttk.Style()
        style.theme_use('clam')

        main_frame = ttk.Frame(self.root, padding=15)
        main_frame.pack(fill='both', expand=True)

        ttk.Label(main_frame, text="유해·위험요인 조사표 엑셀 자동 생성",
                  font=('Malgun Gothic', 15, 'bold')).pack(pady=(0, 10))

        # ── 공통 정보 ──
        info_frame = ttk.LabelFrame(main_frame, text="공통 정보", padding=12)
        info_frame.pack(fill='x', pady=5)

        config = load_config()

        ttk.Label(info_frame, text="수행 일시:").grid(row=0, column=0, sticky='e', padx=5, pady=4)
        self.ent_date = ttk.Entry(info_frame, width=20)
        self.ent_date.insert(0, config.get("eval_date", datetime.now().strftime("%Y.%m.%d")))
        self.ent_date.grid(row=0, column=1, sticky='w', padx=5, pady=4)

        ttk.Label(info_frame, text="수행자 성명:").grid(row=1, column=0, sticky='e', padx=5, pady=4)
        wf = ttk.Frame(info_frame)
        wf.grid(row=1, column=1, sticky='w')
        self.ent_w1 = ttk.Entry(wf, width=10); self.ent_w1.insert(0, config.get("worker1","주진철")); self.ent_w1.pack(side='left', padx=4)
        self.ent_w2 = ttk.Entry(wf, width=10); self.ent_w2.insert(0, config.get("worker2","")); self.ent_w2.pack(side='left', padx=4)
        self.ent_w3 = ttk.Entry(wf, width=10); self.ent_w3.insert(0, config.get("worker3","")); self.ent_w3.pack(side='left', padx=4)

        ttk.Label(info_frame, text="감독 성명:").grid(row=2, column=0, sticky='e', padx=5, pady=4)
        self.ent_sup = ttk.Entry(info_frame, width=10)
        self.ent_sup.insert(0, config.get("supervisor",""))
        self.ent_sup.grid(row=2, column=1, sticky='w', padx=5, pady=4)

        # ── 탭 노트북 ──
        self.nb = ttk.Notebook(main_frame)
        self.nb.pack(fill='both', expand=True, pady=8)

        self.tabs = {}
        tab_defs = [
            ("utpt", "🔬 UT·PT 작업", DEFAULT_UTPT_LEFT, DEFAULT_UTPT_RIGHT),
            ("rt",   "☢ RT 작업",    DEFAULT_RT_LEFT,   DEFAULT_RT_RIGHT),
        ]
        for key, label, def_left, def_right in tab_defs:
            frame = ttk.Frame(self.nb, padding=8)
            self.nb.add(frame, text=label)

            lf = ttk.LabelFrame(frame, text="유해·위험작업 내용", padding=8)
            lf.pack(side='left', fill='both', expand=True, padx=(0,4))
            txt_l = scrolledtext.ScrolledText(lf, width=38, height=18, font=('Malgun Gothic', 9))
            txt_l.pack(fill='both', expand=True)
            txt_l.insert('1.0', config.get(f"{key}_left", def_left))

            rf = ttk.LabelFrame(frame, text="사고, 질병의 유형", padding=8)
            rf.pack(side='right', fill='both', expand=True, padx=(4,0))
            txt_r = scrolledtext.ScrolledText(rf, width=18, height=18, font=('Malgun Gothic', 9))
            txt_r.pack(fill='both', expand=True)
            txt_r.insert('1.0', config.get(f"{key}_right", def_right))

            self.tabs[key] = (txt_l, txt_r)

        # ── 버튼 ──
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=6)
        tk.Button(btn_frame, text="엑셀 파일 생성 (UT·PT + RT 2장)",
                  bg='#1565C0', fg='white', font=('Malgun Gothic', 12, 'bold'),
                  command=self.generate).pack(side='right', ipadx=18, ipady=5)

    def _collect_data(self):
        txt_utpt_l, txt_utpt_r = self.tabs["utpt"]
        txt_rt_l,   txt_rt_r   = self.tabs["rt"]
        return {
            "eval_date":  self.ent_date.get().strip(),
            "worker1":    self.ent_w1.get().strip(),
            "worker2":    self.ent_w2.get().strip(),
            "worker3":    self.ent_w3.get().strip(),
            "supervisor": self.ent_sup.get().strip(),
            "workers":   [self.ent_w1.get().strip(), self.ent_w2.get().strip(), self.ent_w3.get().strip()],
            "utpt": {
                "left_text":  txt_utpt_l.get('1.0','end').strip(),
                "right_text": txt_utpt_r.get('1.0','end').strip(),
            },
            "rt": {
                "left_text":  txt_rt_l.get('1.0','end').strip(),
                "right_text": txt_rt_r.get('1.0','end').strip(),
            },
            "utpt_left":  txt_utpt_l.get('1.0','end').strip(),
            "utpt_right": txt_utpt_r.get('1.0','end').strip(),
            "rt_left":    txt_rt_l.get('1.0','end').strip(),
            "rt_right":   txt_rt_r.get('1.0','end').strip(),
        }

    def generate(self):
        data = self._collect_data()
        save_config(data)

        fname = f"유해위험요인_조사표_{data['eval_date'].replace('.','')}.xlsx"
        output = filedialog.asksaveasfilename(
            defaultextension=".xlsx", initialfile=fname,
            title="저장할 파일 선택", filetypes=[("Excel Files","*.xlsx")])
        if not output:
            return

        success, msg = create_excel(output, data)
        if success:
            messagebox.showinfo("성공", msg)
            try: os.startfile(output)
            except: pass
        else:
            messagebox.showerror("오류", msg)


if __name__ == "__main__":
    root = tk.Tk()
    HazardApp(root)
    root.mainloop()
