import os
import openpyxl

desktop = r'C:\Users\-\OneDrive\바탕 화면'
files = [f for f in os.listdir(desktop) if '산출내역서' in f and f.endswith('.xlsx') and not f.startswith('~$')]

if files:
    filepath = os.path.join(desktop, files[0])
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    print("\n[UT Summary rows]")
    ws = wb['6-2. UT']
    for r in range(1, ws.max_row + 1):
        row_str = "".join([str(ws.cell(row=r, column=c).value) for c in range(1, 4) if ws.cell(row=r, column=c).value is not None])
        if '소계' in row_str or '합계' in row_str or '총계' in row_str or ws.cell(row=r, column=2).value == '계':
            vals = [str(ws.cell(row=r, column=c).value) for c in range(1, 20)]
            print(f"Row {r}: {vals}")

    print("\n[PT Summary rows]")
    ws = wb['6-3. PT']
    for r in range(1, ws.max_row + 1):
        row_str = "".join([str(ws.cell(row=r, column=c).value) for c in range(1, 4) if ws.cell(row=r, column=c).value is not None])
        if '소계' in row_str or '합계' in row_str or '총계' in row_str or ws.cell(row=r, column=2).value == '계':
            vals = [str(ws.cell(row=r, column=c).value) for c in range(1, 20)]
            print(f"Row {r}: {vals}")
