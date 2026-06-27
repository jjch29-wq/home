import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

output_path = r"C:\Users\-\OneDrive\바탕 화면\4.4.1_위험성평가표(RT)_완전체.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "합동 위험성평가표"

# Styles
bold_font = Font(bold=True)
title_font = Font(bold=True, size=18)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

def set_border(ws, min_col, min_row, max_col, max_row):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border

# Columns width setup
widths = [10, 8, 8, 8, 6, 6, 9, 9, 9, 6, 6, 9, 9, 9, 9, 9]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

# --- Title ---
ws.merge_cells('A1:P1')
ws['A1'] = "합동 위험성평가표 [ ■최초, □정기, □수시 ]"
ws['A1'].font = title_font
ws['A1'].alignment = center_align
ws.row_dimensions[1].height = 40

# --- Header Section (Rows 2 to 5) ---
# Row 2
ws.merge_cells('A2:A3')
ws['A2'] = "작 업 명"
ws.merge_cells('B2:D3')
ws['B2'] = "가산~가평 천연가스 공급설비\n건설공사 비파괴검사 기술용역"
ws.merge_cells('E2:F2')
ws['E2'] = "평가기간"
ws.merge_cells('G2:I2')
ws['G2'] = "2026. 07. 09"
ws.merge_cells('J2:K2')
ws['J2'] = "재평가일"
ws.merge_cells('L2:N2')
ws['L2'] = ""
ws.merge_cells('O2:P2')

# Row 3 (Company)
ws.merge_cells('A4:A5')
ws['A4'] = "회 사 명"
ws.merge_cells('B4:D5')
ws['B4'] = "서울검사(주)"

# Row 4 (Process)
ws.merge_cells('A6:A7')
ws['A6'] = "작업공종"
ws.merge_cells('B6:D7')
ws['B6'] = "방사선투과검사"

# Row 5 (Period)
ws.merge_cells('A8:A9')
ws['A8'] = "작업기간"
ws.merge_cells('B8:D9')
ws['B8'] = "2026.06.22 ~ 2028.10.31"

# Signatures Block (작업전)
ws.merge_cells('E3:E9')
ws['E3'] = "작업전"

ws.merge_cells('F3:F5')
ws['F3'] = "수급인"
ws['G3'] = "근 로 자"
ws['H3'] = "작 성 자"
ws['I3'] = "승 인 자"
ws.merge_cells('G4:G5')
ws['G4'] = "유상훈 (서명)"
ws.merge_cells('H4:H5')
ws['H4'] = "주진철 (서명)"
ws.merge_cells('I4:I5')
ws['I4'] = "강신태 (서명)"

ws.merge_cells('F6:F9')
ws['F6'] = "도급인"
ws['G6'] = "검토자(감독)"
ws['H6'] = "검토자(안전)"
ws['I6'] = "승 인 자"
ws.merge_cells('G7:G9')
ws['G7'] = "(서명)"
ws.merge_cells('H7:H9')
ws['H7'] = "(서명)"
ws.merge_cells('I7:I9')
ws['I7'] = "(서명)"

# Signatures Block (재평가)
ws.merge_cells('J3:J9')
ws['J3'] = "재평가"

ws.merge_cells('K3:K5')
ws['K3'] = "수급인"
ws['L3'] = "근 로 자"
ws['M3'] = "작 성 자"
ws['N3'] = "승 인 자"
ws.merge_cells('L4:L5')
ws['L4'] = "(서명)"
ws.merge_cells('M4:M5')
ws['M4'] = "(서명)"
ws.merge_cells('N4:N5')
ws['N4'] = "(서명)"

ws.merge_cells('K6:K9')
ws['K6'] = "도급인"
ws['L6'] = "검토자(감독)"
ws['M6'] = "검토자(안전)"
ws['N6'] = "승 인 자"
ws.merge_cells('L7:L9')
ws['L7'] = "(서명)"
ws.merge_cells('M7:M9')
ws['M7'] = "(서명)"
ws.merge_cells('N7:N9')
ws['N7'] = "(서명)"

# Clean up empty P col from headers
ws.merge_cells('O3:P9')

for r in range(2, 10):
    for c in range(1, 15):
        cell = ws.cell(row=r, column=c)
        cell.alignment = center_align
        cell.font = Font(size=10)

set_border(ws, 1, 2, 14, 9)

# --- Table Sub-Header ---
ws.merge_cells('A10:I10')
ws['A10'] = "작업 전 평가"
ws.merge_cells('J10:N10')
ws['J10'] = "재평가"

ws.merge_cells('A11:A12')
ws['A11'] = "세부작업"
ws.merge_cells('B11:D11')
ws['B11'] = "유해 위험 요인 파악"
ws.merge_cells('E11:E12')
ws['E11'] = "현재의 안전 보건 조치"
ws.merge_cells('F11:H11')
ws['F11'] = "위험성 결정"
ws.merge_cells('I11:I12')
ws['I11'] = "위험성 감소 대책"
ws.merge_cells('J11:J12')
ws['J11'] = "개선\n예정일"
ws.merge_cells('K11:K12')
ws['K11'] = "담당자\n(담당부서)"
ws.merge_cells('L11:L12')
ws['L11'] = "완료일"
ws.merge_cells('M11:N11')
ws['M11'] = "개선 후 위험성 (재평가)"

ws['B12'] = "위험\n분류"
ws.merge_cells('C12:D12')
ws['C12'] = "위험 발생 상황 및 결과"
ws['F12'] = "가능성\n(빈도)"
ws['G12'] = "중대성\n(강도)"
ws['H12'] = "위험성\n(등급)"
ws['M12'] = "가능성\n(빈도)"
ws['N12'] = "중대성\n(강도)"
# (Skip writing P here)

for r in range(10, 13):
    for c in range(1, 15):
        cell = ws.cell(row=r, column=c)
        cell.alignment = center_align
        cell.font = bold_font
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

set_border(ws, 1, 10, 14, 12)

# --- Data Rows ---
data = [
    ["필름부착 작업 시 배관상부에서 미끄러져 넘어짐 위험", "안전모, 안전화 착용 / 작업 전 점검", 3, 2, "V", "1. 2m 이상 고소작업 시 지정된 생명줄에 안전대(그네식) 체결 필수\n2. 우천/결빙 시 배관 상부 탑승 작업 전면 중지"],
    ["지하 작업 중 낙하물에 의한 사고위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 굴착 상단부에 자재 및 공구 적재 금지\n2. 하부 작업 시 상부 타 공정 동시 작업 원칙적 금지"],
    ["필름부착작업 시 사다리 이동작업 중 추락/전도 위험", "안전모, 안전화 착용 / 작업 전 점검", 3, 2, "V", "1. A형 사다리 사용 시 반드시 2인 1조 작업\n2. 사다리 아웃트리거 전개 및 최상단 발판 탑승 금지"],
    ["차폐체 설치 작업 시 낙하에 의한 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 100kg 이상 중량물 인양 시 반드시 2줄 걸이 이상 결속\n2. 인양 중인 중량물 하부 통행 전면 금지"],
    ["차폐체설치 체인블럭 고정상태 불량시 낙하에 의한 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 인양 전 체인블록, 훅(Hook) 마모/파손 상태 점검\n2. 고정 지지대 하중 강도 확인 후 거치"],
    ["손상된 슬링벨트 사용 시 판단에 의한 중량물낙하 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 작업 전 슬링벨트 손상 및 규격 확인\n2. 파손된 슬링벨트는 즉시 파기 및 교체"],
    ["중량물 줄걸이작업 부적합 시 낙하에 의한 사고위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 지정된 줄걸이 용구 사용 및 2줄걸이 이상 작업\n2. 관리감독자 지휘 하에 안전하게 인양 진행"]
]

start_row = 13
for i, row_data in enumerate(data):
    r = start_row + i
    
    # 세부작업, 위험분류 (Merged for all rows)
    if i == 0:
        ws.cell(row=r, column=1, value="사전준비").alignment = center_align
        ws.cell(row=r, column=2, value="기계적\n(설비)\n요인").alignment = center_align
    
    # Data columns
    ws.merge_cells(f'C{r}:D{r}')
    ws.cell(row=r, column=3, value=row_data[0]).alignment = left_align
    ws.cell(row=r, column=5, value=row_data[1]).alignment = left_align
    ws.cell(row=r, column=6, value=row_data[2]).alignment = center_align
    ws.cell(row=r, column=7, value=row_data[3]).alignment = center_align
    ws.cell(row=r, column=8, value=row_data[4]).alignment = center_align
    
    # 대책 column
    ws.cell(row=r, column=9, value=row_data[5]).alignment = left_align
    
    ws.row_dimensions[r].height = 45

ws.merge_cells(f'A13:A{start_row+len(data)-1}')
ws.merge_cells(f'B13:B{start_row+len(data)-1}')

set_border(ws, 1, 13, 14, start_row+len(data)-1)

wb.save(output_path)
print(f"Excel file created at: {output_path}")
