import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import os
from datetime import datetime, timedelta
import json

CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config_meeting.json")

class PreTrainingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("위험성평가 사전교육(회의) 자동 생성기")
        self.root.geometry("650x1000")
        
        style = ttk.Style()
        style.theme_use('clam')
        
        main_frame = ttk.Frame(self.root, padding=20)
        main_frame.pack(fill='both', expand=True)
        
        ttk.Label(main_frame, text="사전교육(회의)록 자동 생성", font=('Malgun Gothic', 16, 'bold')).pack(pady=(0, 20))
        
        form_frame = ttk.LabelFrame(main_frame, text="입력 정보", padding=15)
        form_frame.pack(fill='x', pady=5)
        
        ttk.Label(form_frame, text="사전교육 일시(붙임1):").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.ent_date = ttk.Entry(form_frame, width=55)
        
        now = datetime.now()
        end_time = now + timedelta(hours=1)
        default_time_str = f"{now.strftime('%Y년 %m월 %d일 %H:%M')} ~ {end_time.strftime('%H:%M')} (1시간)"
        
        self.ent_date.insert(0, default_time_str)
        self.ent_date.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(form_frame, text="전파교육 일시(붙임2):").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.ent_date2 = ttk.Entry(form_frame, width=55)
        dissem_start = now + timedelta(hours=5)
        dissem_end = dissem_start + timedelta(hours=1)
        default_time_str2 = f"{dissem_start.strftime('%Y년 %m월 %d일 %H:%M')} ~ {dissem_end.strftime('%H:%M')} (1시간)"
        self.ent_date2.insert(0, default_time_str2)
        self.ent_date2.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(form_frame, text="사전교육 장소(붙임1):").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        self.ent_loc = ttk.Entry(form_frame, width=55)
        self.ent_loc.insert(0, "인천사무실")
        self.ent_loc.grid(row=2, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(form_frame, text="전파교육 장소(붙임2):").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        self.ent_loc2 = ttk.Entry(form_frame, width=55)
        self.ent_loc2.insert(0, "가산~가평 현장 안전교육장")
        self.ent_loc2.grid(row=3, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(form_frame, text="평가 리더:").grid(row=4, column=0, sticky='e', padx=5, pady=2)
        self.ent_leader = ttk.Entry(form_frame, width=55)
        self.ent_leader.insert(0, "곽재운 이사(현장소장)")
        self.ent_leader.grid(row=4, column=1, sticky='w', padx=5, pady=2)
        
        ttk.Label(form_frame, text="관리감독자:").grid(row=5, column=0, sticky='e', padx=5, pady=2)
        self.ent_super = ttk.Entry(form_frame, width=55)
        self.ent_super.insert(0, "주진철")
        self.ent_super.grid(row=5, column=1, sticky='w', padx=5, pady=2)
        
        ttk.Label(form_frame, text="근로자대표:").grid(row=6, column=0, sticky='e', padx=5, pady=2)
        self.ent_worker = ttk.Entry(form_frame, width=55)
        self.ent_worker.insert(0, "유상훈")
        self.ent_worker.grid(row=6, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(form_frame, text="참석자:").grid(row=7, column=0, sticky='e', padx=5, pady=5)
        self.txt_attendees = tk.Text(form_frame, width=60, height=3, font=('Malgun Gothic', 9))
        self.txt_attendees.insert('1.0', "유상훈(서울검사/사원), 주진철(서울검사/팀장), 강신태(서울검사/소장)")
        self.txt_attendees.grid(row=7, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(form_frame, text="(예시: 이름(소속/직책) 형식으로 쉼표/엔터 구분)", foreground="gray", font=("Malgun Gothic", 8)).grid(row=8, column=1, sticky='w', padx=5)
        
        ttk.Label(form_frame, text="회의 내용:").grid(row=9, column=0, sticky='ne', padx=5, pady=5)
        self.txt_content = tk.Text(form_frame, width=60, height=9, font=('Malgun Gothic', 9))
        default_content = (
            "1. 위험성평가의 목적 및 방법, 시기 및 절차\n"
            "[2. 평가담당자 및 책임자의 역할 - 생성 시 자동 삽입]\n"
            "3. 근로자에 대한 참여·공유방법 및 유의사항\n"
            "4. 위험성평가 결과의 기록·보존 방법\n"
            "5. 금회 위험성평가 실시 대상 공정\n"
            " - 방사선투과(RT), 초음파(UT), 침투(PT) 및 가설컨테이너 운영\n"
            "6. 중점 관리(논의) 사항\n"
            " - [방사선] 야간 RT 검사 시 타 공정 근로자 출입 통제구역(10μSv/hr) 확보\n"
            " - [낙하] 슬링벨트 훼손품 즉시 폐기 및 소형장비 가방 운반 원칙\n"
            " - [전기] 우천/야간 작업 시 감전 예방 위해 누전차단기 부착 릴선 전용 사용"
        )
        self.txt_content.insert('1.0', default_content)
        self.txt_content.grid(row=9, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(form_frame, text="전파교육 내용:").grid(row=10, column=0, sticky='ne', padx=5, pady=5)
        self.txt_content2 = tk.Text(form_frame, width=60, height=5, font=('Malgun Gothic', 9))
        default_content2 = (
            "1. 해당 작업과 관련된 유해·위험요인 및 위험성 결정 결과 주지\n"
            "2. 유해·위험요인에 대한 위험성 감소대책과 실행 계획\n"
            "3. 위험성 감소대책에 따라 근로자가 준수하거나 주의하여야 할 사항\n"
            "4. [현장 주요 전파사항] 방사선 통제구역 출입금지, 노후 슬링벨트 사용금지, 누전차단기 전용 릴선 사용 준수"
        )
        self.txt_content2.insert('1.0', default_content2)
        self.txt_content2.grid(row=10, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(form_frame, text="회의 사진(붙임1):").grid(row=11, column=0, sticky='e', padx=5, pady=5)
        photo_frame1 = ttk.Frame(form_frame)
        photo_frame1.grid(row=11, column=1, sticky='w', padx=5, pady=5)
        
        self.photo_path1 = None
        self.lbl_photo1 = ttk.Label(photo_frame1, text="선택된 사진 없음", foreground="gray", width=25)
        self.lbl_photo1.pack(side='left', padx=(0, 5))
        ttk.Button(photo_frame1, text="사진 찾기", command=self.select_photo1).pack(side='left')
        
        ttk.Label(form_frame, text="전파교육 사진(붙임2):").grid(row=12, column=0, sticky='e', padx=5, pady=5)
        photo_frame2 = ttk.Frame(form_frame)
        photo_frame2.grid(row=12, column=1, sticky='w', padx=5, pady=5)
        
        self.photo_path2 = None
        self.lbl_photo2 = ttk.Label(photo_frame2, text="선택된 사진 없음", foreground="gray", width=25)
        self.lbl_photo2.pack(side='left', padx=(0, 5))
        ttk.Button(photo_frame2, text="사진 찾기", command=self.select_photo2).pack(side='left')
        
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=20)
        
        self.btn_generate = ttk.Button(btn_frame, text="엑셀 파일 생성", command=self.generate_excel)
        self.btn_generate.pack(fill='x', ipady=5)
        
        self.lbl_status = ttk.Label(main_frame, text="대기 중...", foreground="gray")
        self.lbl_status.pack()

        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.load_config()

    def save_config(self):
        config = {
            "date": self.ent_date.get(),
            "date2": self.ent_date2.get(),
            "loc": self.ent_loc.get(),
            "loc2": self.ent_loc2.get(),
            "leader": self.ent_leader.get(),
            "superv": self.ent_super.get(),
            "worker": self.ent_worker.get(),
            "attendees": self.txt_attendees.get("1.0", tk.END).strip(),
            "content": self.txt_content.get("1.0", tk.END).strip(),
            "content2": self.txt_content2.get("1.0", tk.END).strip()
        }
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(config, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print("설정 저장 실패:", e)

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    config = json.load(f)
                
                if "date" in config:
                    self.ent_date.delete(0, tk.END)
                    self.ent_date.insert(0, config["date"])
                if "date2" in config:
                    self.ent_date2.delete(0, tk.END)
                    self.ent_date2.insert(0, config["date2"])
                if "loc" in config:
                    self.ent_loc.delete(0, tk.END)
                    self.ent_loc.insert(0, config["loc"])
                if "loc2" in config:
                    self.ent_loc2.delete(0, tk.END)
                    self.ent_loc2.insert(0, config["loc2"])
                if "leader" in config:
                    self.ent_leader.delete(0, tk.END)
                    self.ent_leader.insert(0, config["leader"])
                if "superv" in config:
                    self.ent_super.delete(0, tk.END)
                    self.ent_super.insert(0, config["superv"])
                if "worker" in config:
                    self.ent_worker.delete(0, tk.END)
                    self.ent_worker.insert(0, config["worker"])
                if "attendees" in config:
                    self.txt_attendees.delete("1.0", tk.END)
                    self.txt_attendees.insert("1.0", config["attendees"])
                if "content" in config:
                    self.txt_content.delete("1.0", tk.END)
                    self.txt_content.insert("1.0", config["content"])
                if "content2" in config:
                    self.txt_content2.delete("1.0", tk.END)
                    self.txt_content2.insert("1.0", config["content2"])
            except Exception as e:
                print("설정 불러오기 실패:", e)

    def on_closing(self):
        self.save_config()
        self.root.destroy()

    def select_photo1(self):
        paths = filedialog.askopenfilenames(filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp")])
        if paths:
            self.photo_path1 = paths
            if len(paths) == 1:
                self.lbl_photo1.config(text=os.path.basename(paths[0]), foreground="black")
            else:
                self.lbl_photo1.config(text=f"{len(paths)}장 선택됨", foreground="black")

    def select_photo2(self):
        paths = filedialog.askopenfilenames(filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp")])
        if paths:
            self.photo_path2 = paths
            if len(paths) == 1:
                self.lbl_photo2.config(text=os.path.basename(paths[0]), foreground="black")
            else:
                self.lbl_photo2.config(text=f"{len(paths)}장 선택됨", foreground="black")

    def generate_excel(self):
        # 엑셀 생성 시 현재 설정(참석자 등)을 자동 저장합니다.
        self.save_config()
        
        try:
            self.btn_generate.config(state='disabled')
            self.lbl_status.config(text="생성 중...", foreground="blue")
            self.root.update()
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "사전교육회의록"
            
            # --- 1. 페이지 설정 (A4 가득 차게) ---
            ws.page_setup.orientation = 'portrait'
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0 # 세로는 내용에 맞춤
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.6
            ws.page_margins.bottom = 0.6
            ws.print_options.horizontalCentered = True
            
            # --- 2. 열 너비 설정 (A~F) ---
            # 비율: 13, 14, 18, 13, 14, 18 (총 90 언저리)
            widths = [13, 15, 18, 13, 15, 18]
            for i, width in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
                
            # --- 3. 스타일 세팅 ---
            bold_font = Font(name='맑은 고딕', size=11, bold=True)
            normal_font = Font(name='맑은 고딕', size=11)
            title_font = Font(name='맑은 고딕', size=20, bold=True)
            blue_font = Font(name='맑은 고딕', size=11, bold=True, color="0000FF")
            fill_gray = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
            
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            thin = Side(style='thin')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            thick = Side(style='thick')
            thick_border_top = Border(top=thick, left=thin, right=thin, bottom=thin)

            def set_border(min_col, min_row, max_col, max_row):
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        ws.cell(row=r, column=c).border = border

            # --- 4. 제목부 ---
            ws['A1'] = "[붙임1]"
            ws['A1'].font = blue_font
            ws['A1'].alignment = left_align
            
            ws.merge_cells('A2:F2')
            ws['A2'] = "위험성평가 사전교육(회의)"
            ws['A2'].font = title_font
            ws['A2'].alignment = center_align
            ws.row_dimensions[2].height = 40
            
            ws.row_dimensions[3].height = 10
            
            # --- 5. 일시 및 장소 테이블 ---
            ws.merge_cells('A4:B4')
            ws['A4'] = "일   시"
            ws['A4'].font = normal_font
            ws['A4'].alignment = center_align
            ws['A4'].fill = fill_gray
            
            ws.merge_cells('C4:F4')
            ws['C4'] = self.ent_date.get()
            ws['C4'].font = normal_font
            ws['C4'].alignment = center_align
            
            ws.merge_cells('A5:B5')
            ws['A5'] = "장   소"
            ws['A5'].font = normal_font
            ws['A5'].alignment = center_align
            ws['A5'].fill = fill_gray
            
            ws.merge_cells('C5:F5')
            ws['C5'] = self.ent_loc.get()
            ws['C5'].font = normal_font
            ws['C5'].alignment = center_align
            
            set_border(1, 4, 6, 5)
            # 표 바깥선 두껍게 (상단)
            for col in range(1, 7):
                ws.cell(row=4, column=col).border = Border(top=thick, left=thin, right=thin, bottom=thin)
            
            ws.row_dimensions[4].height = 25
            ws.row_dimensions[5].height = 25
            
            # --- 6. 교육(회의)내용 체크리스트 ---
            ws.merge_cells('A6:F6')
            ws['A6'] = " □ 교육(회의)내용"
            ws['A6'].font = Font(name='맑은 고딕', size=12, bold=False)
            ws['A6'].alignment = Alignment(horizontal='left', vertical='bottom')
            ws.row_dimensions[6].height = 30
            
            # 동적 텍스트 생성 (UI 필드 + 텍스트 박스)
            leader = self.ent_leader.get()
            superv = self.ent_super.get()
            worker = self.ent_worker.get()
            
            roles_text = f"2. 평가담당자 및 책임자의 역할\n - 평가 리더: {leader} / 관리감독자: {superv} / 근로자 대표: {worker}"
            
            raw_content = self.txt_content.get("1.0", tk.END).strip()
            # 텍스트 박스 내용 중 "[2. 평가담당자 및 책임자의 역할 - 생성 시 자동 삽입]" 부분을 치환
            if "[2. 평가담당자 및 책임자의 역할 - 생성 시 자동 삽입]" in raw_content:
                content_text = raw_content.replace("[2. 평가담당자 및 책임자의 역할 - 생성 시 자동 삽입]", roles_text)
            else:
                # 이전 버전의 내용이 저장된 경우에 대한 호환성 유지
                content_text = ""
                inserted = False
                for line in raw_content.split('\n'):
                    if not inserted and line.startswith("2."):
                        content_text += roles_text + "\n"
                        content_text += line.replace("2.", "3.") + "\n"
                        inserted = True
                    else:
                        content_text += line + "\n"
                        
            content_text = content_text.strip()
            
            ws.merge_cells('A7:F7')
            ws['A7'] = content_text
            ws['A7'].font = normal_font
            ws['A7'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            set_border(1, 7, 6, 7)
            # 상단 굵은선
            for col in range(1, 7):
                ws.cell(row=7, column=col).border = Border(top=thick, left=thin, right=thin, bottom=thin)
            ws.row_dimensions[7].height = 200
            
            # --- 7. 빈 공간 (사진 첨부부) ---
            ws.merge_cells('A8:F8')
            set_border(1, 8, 6, 8)
            ws.row_dimensions[8].height = 230
            
            if self.photo_path1:
                self._insert_photos(ws, self.photo_path1, 'A8')
            
            # --- 8. 참석자 명단 테이블 ---
            ws.merge_cells('A9:F9')
            ws['A9'] = " □ 참석자 명단"
            ws['A9'].font = Font(name='맑은 고딕', size=12, bold=False)
            ws['A9'].alignment = Alignment(horizontal='left', vertical='bottom')
            ws.row_dimensions[9].height = 40
            
            headers = ["소속/직책", "성   명", "서   명", "소속/직책", "성   명", "서   명"]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=10, column=c)
                cell.value = h
                cell.font = normal_font
                cell.alignment = center_align
                cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)
            ws.row_dimensions[10].height = 25
            
            # 입력받은 참석자 파싱 (이름과 괄호 안의 소속/직책 분리)
            import re
            # 줄바꿈으로 입력할 수도 있으므로 쉼표로 치환 후 파싱
            raw_attendees_text = self.txt_attendees.get("1.0", tk.END).strip().replace('\n', ',')
            raw_attendees = raw_attendees_text.split(',')
            attendees_parsed = []
            for raw in raw_attendees:
                raw = raw.strip()
                if not raw: continue
                
                match = re.match(r'([^(]+)(?:\(([^)]+)\))?', raw)
                if match:
                    name = match.group(1).strip()
                    pos = match.group(2).strip() if match.group(2) else ""
                else:
                    name = raw
                    pos = ""
                attendees_parsed.append({"name": name, "pos": pos})
            
            # 2개씩 짝지어서 배열
            pairs = []
            for i in range(0, len(attendees_parsed), 2):
                p1 = attendees_parsed[i]
                p2 = attendees_parsed[i+1] if i+1 < len(attendees_parsed) else {"name": "", "pos": ""}
                pairs.append((p1, p2))
                
            # 최소 4줄 이상 생성 (양식의 통일감을 위해)
            while len(pairs) < 4:
                pairs.append(({"name": "", "pos": ""}, {"name": "", "pos": ""}))
                
            current_row = 11
            for p1, p2 in pairs:
                for col in range(1, 7):
                    ws.cell(row=current_row, column=col).border = border
                    ws.cell(row=current_row, column=col).font = normal_font
                    ws.cell(row=current_row, column=col).alignment = center_align
                    
                ws.cell(row=current_row, column=1).value = p1["pos"]
                ws.cell(row=current_row, column=2).value = p1["name"]
                ws.cell(row=current_row, column=4).value = p2["pos"]
                ws.cell(row=current_row, column=5).value = p2["name"]
                ws.row_dimensions[current_row].height = 35
                
                # 도장 이미지 삽입 함수 호출
                self.insert_signature(ws, p1["name"], current_row, 'C')
                self.insert_signature(ws, p2["name"], current_row, 'F')
                
                current_row += 1
                
            # ==========================================
            # 시트 2: 위험성평가 결과 전파교육 생성
            # ==========================================
            ws2 = wb.create_sheet(title="전파교육")
            
            # 1. 페이지 설정
            ws2.page_setup.orientation = 'portrait'
            ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
            ws2.page_setup.fitToPage = True
            ws2.page_setup.fitToWidth = 1
            ws2.page_setup.fitToHeight = 0
            ws2.page_margins.left = 0.5
            ws2.page_margins.right = 0.5
            ws2.page_margins.top = 0.6
            ws2.page_margins.bottom = 0.6
            ws2.print_options.horizontalCentered = True
            
            # 2. 열 너비 설정
            for i, width in enumerate(widths, 1):
                ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
                
            def set_border2(min_col, min_row, max_col, max_row):
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        ws2.cell(row=r, column=c).border = border

            # 3. 제목부
            ws2['A1'] = "[붙임2]"
            ws2['A1'].font = blue_font
            ws2['A1'].alignment = left_align
            
            ws2.merge_cells('A2:F2')
            ws2['A2'] = "위험성평가 결과 전파교육"
            ws2['A2'].font = title_font
            ws2['A2'].alignment = center_align
            ws2.row_dimensions[2].height = 40
            
            ws2.row_dimensions[3].height = 10
            
            # 4. 일시 및 장소 (시트 1과 동일)
            ws2.merge_cells('A4:B4')
            ws2['A4'] = "일   시"
            ws2['A4'].font = normal_font
            ws2['A4'].alignment = center_align
            ws2['A4'].fill = fill_gray
            
            ws2.merge_cells('C4:F4')
            ws2['C4'] = self.ent_date2.get()
            ws2['C4'].font = normal_font
            ws2['C4'].alignment = center_align
            
            ws2.merge_cells('A5:B5')
            ws2['A5'] = "장   소"
            ws2['A5'].font = normal_font
            ws2['A5'].alignment = center_align
            ws2['A5'].fill = fill_gray
            
            ws2.merge_cells('C5:F5')
            ws2['C5'] = self.ent_loc2.get()
            ws2['C5'].font = normal_font
            ws2['C5'].alignment = center_align
            
            set_border2(1, 4, 6, 5)
            for col in range(1, 7):
                ws2.cell(row=4, column=col).border = Border(top=thick, left=thin, right=thin, bottom=thin)
            ws2.row_dimensions[4].height = 25
            ws2.row_dimensions[5].height = 25
            
            # 5. 교육내용
            ws2.merge_cells('A6:F6')
            ws2['A6'] = " □ 교육내용"
            ws2['A6'].font = Font(name='맑은 고딕', size=12, bold=False)
            ws2['A6'].alignment = Alignment(horizontal='left', vertical='bottom')
            ws2.row_dimensions[6].height = 30
            
            edu_content = self.txt_content2.get("1.0", tk.END).strip()
            ws2.merge_cells('A7:F7')
            ws2['A7'] = edu_content
            ws2['A7'].font = normal_font
            ws2['A7'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            set_border2(1, 7, 6, 7)
            for col in range(1, 7):
                ws2.cell(row=7, column=col).border = Border(top=thick, left=thin, right=thin, bottom=thin)
            ws2.row_dimensions[7].height = 200
            
            # 6. 빈 공간 (사진 첨부부 - 전파교육용 사진)
            ws2.merge_cells('A8:F8')
            set_border2(1, 8, 6, 8)
            ws2.row_dimensions[8].height = 230
            
            if self.photo_path2:
                self._insert_photos(ws2, self.photo_path2, 'A8')
                    
            # 7. 참석자 명단 테이블
            ws2.merge_cells('A9:F9')
            ws2['A9'] = " □ 참석자 명단"
            ws2['A9'].font = Font(name='맑은 고딕', size=12, bold=False)
            ws2['A9'].alignment = Alignment(horizontal='left', vertical='bottom')
            ws2.row_dimensions[9].height = 40
            
            for c, h in enumerate(headers, 1):
                cell = ws2.cell(row=10, column=c)
                cell.value = h
                cell.font = normal_font
                cell.alignment = center_align
                cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)
            ws2.row_dimensions[10].height = 25
            
            # 참석자 목록 (시트 1과 동일한 pairs 재사용)
            current_row = 11
            for p1, p2 in pairs:
                for col in range(1, 7):
                    ws2.cell(row=current_row, column=col).border = border
                    ws2.cell(row=current_row, column=col).font = normal_font
                    ws2.cell(row=current_row, column=col).alignment = center_align
                    
                ws2.cell(row=current_row, column=1).value = p1["pos"]
                ws2.cell(row=current_row, column=2).value = p1["name"]
                ws2.cell(row=current_row, column=4).value = p2["pos"]
                ws2.cell(row=current_row, column=5).value = p2["name"]
                ws2.row_dimensions[current_row].height = 35
                
                self.insert_signature(ws2, p1["name"], current_row, 'C')
                self.insert_signature(ws2, p2["name"], current_row, 'F')
                
                current_row += 1
                
            # --- 9. 엑셀 파일 저장 ---
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=f"사전교육회의록_{datetime.now().strftime('%Y%m%d')}.xlsx",
                title="엑셀 파일 저장",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if save_path:
                wb.save(save_path)
                # 엑셀 파일 생성 성공 시 설정도 저장
                self.save_config()
                messagebox.showinfo("완료", "엑셀 파일이 성공적으로 생성되었습니다!")
                self.lbl_status.config(text="생성 완료!", foreground="green")
            else:
                self.lbl_status.config(text="저장 취소됨", foreground="gray")
                
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 생성 중 오류 발생:\n{e}")
            self.lbl_status.config(text="오류 발생", foreground="red")
        finally:
            self.btn_generate.config(state='normal')

    def _insert_photos(self, ws, paths, cell):
        if not paths: return
        if isinstance(paths, str): paths = [paths]
        valid_paths = [p for p in paths if os.path.exists(p)]
        if not valid_paths: return
        
        import math, io
        canvas_width, canvas_height = 692, 373
        margin = 6
        inner_width = canvas_width - margin * 2
        inner_height = canvas_height - margin * 2
        num_images = len(valid_paths)
        
        canvas = PILImage.new('RGBA', (canvas_width, canvas_height), (255, 255, 255, 0))
        
        if num_images == 1:
            try:
                with PILImage.open(valid_paths[0]) as img:
                    img = img.convert("RGBA")
                    w, h = img.size
                    ratio = min(inner_width/w, inner_height/h)
                    new_w, new_h = int(w * ratio), int(h * ratio)
                    img_resized = img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
                    
                    # 상하좌우 중앙 정렬 (마진 포함한 캔버스 기준)
                    x = (canvas_width - new_w) // 2
                    y = (canvas_height - new_h) // 2
                    canvas.paste(img_resized, (x, y), img_resized)
            except Exception as ex:
                print("사진 삽입 실패:", ex)
        else:
            try:
                cols = math.ceil(math.sqrt(num_images))
                rows = math.ceil(num_images / cols)
                if num_images in (2, 3):
                    cols, rows = num_images, 1
                    
                box_w, box_h = inner_width // cols, inner_height // rows
                
                for i, path in enumerate(valid_paths):
                    row_idx, col_idx = i // cols, i % cols
                    try:
                        with PILImage.open(path) as img:
                            img = img.convert("RGBA")
                            w, h = img.size
                            gap = 10
                            avail_w, avail_h = box_w - gap, box_h - gap
                            
                            # 원본 비율을 유지하며 축소
                            ratio = min(avail_w/w, avail_h/h)
                            new_w, new_h = int(w * ratio), int(h * ratio)
                            img_resized = img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
                            
                            # 각 영역의 상하좌우 중앙 정렬 (마진 오프셋 추가)
                            x = margin + col_idx * box_w + (box_w - new_w) // 2
                            y = margin + row_idx * box_h + (box_h - new_h) // 2
                            canvas.paste(img_resized, (x, y), img_resized)
                    except Exception as e:
                        print(f"이미지 처리 실패 {path}: {e}")
            except Exception as ex:
                print("사진 그리드 삽입 실패:", ex)
                
        try:
            img_byte_arr = io.BytesIO()
            canvas.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            xl_img = XLImage(img_byte_arr)
            ws.add_image(xl_img, cell)
        except Exception as ex:
            print("사진 엑셀 삽입 실패:", ex)

    def insert_signature(self, ws, name, row, col_letter):
        if not name: return
        
        possible_names = [f"{name}_padded.png", f"{name}.png"]
        base_dir = os.path.dirname(os.path.abspath(__file__))
        signs_dir = os.path.join(base_dir, "signs")
        
        found_path = None
        for n in possible_names:
            p = os.path.join(signs_dir, n)
            if os.path.exists(p):
                found_path = p
                break
                
        if found_path:
            try:
                import io
                # 셀 크기: 너비 18(약 126px), 높이 35(약 46px)
                # 투명한 배경 캔버스 생성 (셀 크기에 맞춤)
                canvas = PILImage.new('RGBA', (120, 45), (255, 255, 255, 0))
                with PILImage.open(found_path) as sign_img:
                    # 원본 도장을 적당한 크기(50x35)로 리사이즈
                    sign_img = sign_img.convert("RGBA")
                    sign_img = sign_img.resize((50, 35), PILImage.Resampling.LANCZOS)
                    
                    # 캔버스 정중앙에 도장 붙이기
                    # 가로 중앙: (120 - 50) / 2 = 35
                    # 세로 중앙: (45 - 35) / 2 = 5
                    canvas.paste(sign_img, (35, 5), sign_img)
                
                # 메모리에 이미지 저장 후 엑셀에 삽입
                img_byte_arr = io.BytesIO()
                canvas.save(img_byte_arr, format='PNG')
                img_byte_arr.seek(0)
                
                xl_img = XLImage(img_byte_arr)
                ws.add_image(xl_img, f"{col_letter}{row}")
            except Exception as e:
                print("서명 삽입 실패:", e)

if __name__ == "__main__":
    root = tk.Tk()
    app = PreTrainingApp(root)
    root.mainloop()
