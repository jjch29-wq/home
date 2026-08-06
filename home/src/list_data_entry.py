# -*- coding: utf-8 -*-
"""
⚡ 종합 리스트 초고속 입력기
대상 파일: 검사_LIST_중앙지사_서식완성.xlsx
대상 시트: 중앙지사작업

컬럼 구조 (0-indexed):
 C0:순번  C1:촬영일자  C2:공정  C3:Section  C4:Line No.  C5:Joint
 C6:용접사번호  C7~C10:촬영구간1~4  C11:관경  C12:규격  C13:합부
 C14:RT촬영매수ORI  C15:RT촬영매수RE'
 C16:PAUT길이ORI  C17:PAUT길이RE'
 C18:MT길이  C19:기성  C20:비고
"""
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
import os
import re
import datetime

try:
    from tkcalendar import DateEntry
except ImportError:
    DateEntry = None

# ── 대상 파일 경로 ──
EXCEL_PATH = r'C:\Users\-\OneDrive\바탕 화면\검사_LIST_중앙지사_서식완성.xlsx'
SHEET_NAME = '중앙지사작업'

# 헤더는 행 1~3 (0-indexed 기준 R0~R2) → 데이터는 행 4부터 (openpyxl row=4)
DATA_START_ROW = 4  # openpyxl 1-indexed

# 공정별 활성 컬럼 정의 (openpyxl 1-indexed)
COL = {
    '순번':         1,
    '촬영일자':     2,
    '공정':         3,
    'Section':      4,
    'Line No.':     5,
    'Joint':        6,
    '용접사번호':   7,
    '구간1':        8,
    '구간2':        9,
    '구간3':       10,
    '구간4':       11,
    '관경':        12,
    '규격':        13,
    '합부':        14,
    'RT_ORI':      15,
    "RT_RE'":      16,
    'PAUT_ORI':    17,
    "PAUT_RE'":    18,
    'MT길이':      19,
    '기성':        20,
    '비고':        21,
}
MAX_COL = 21

# 크기 목록
SIZES = ['1100A','1000A','900A','850A','800A','750A','700A','650A',
         '600A','550A','500A','450A','400A','350A','300A','250A',
         '200A','150A','125A','100A','80A','65A','50A','40A','32A','25A','20A']

# 관경별 PAUT/MT/PT 검사길이 (m) 룩업 테이블 ← 검사길이 시트 기준
SIZE_LENGTH = {
    '1100A': 3.511,  '1000A': 3.1919, '900A': 2.8727, '850A': 2.7131,
    '800A':  2.5535, '750A':  2.3939, '700A': 2.2343, '650A': 2.0747,
    '600A':  1.9151, '550A':  1.7555, '500A': 1.5959, '450A': 1.4363,
    '400A':  1.2767, '350A':  1.1172, '300A': 1.0006, '250A': 0.8401,
    '200A':  0.6795, '150A':  0.519,  '125A': 0.4392, '100A': 0.3591,
    '80A':   0.2799, '65A':   0.2397, '50A':  0.1901, '40A':  0.1527,
    '32A':   0.1341, '25A':   0.1068, '20A':  0.0855,
}


class DataEntryUI(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("⚡ 종합 리스트 초고속 입력기")
        self.geometry("860x700")
        self.resizable(True, True)
        self.attributes('-topmost', True)

        self.excel_path = EXCEL_PATH

        # ── 공통 변수 ──
        self.process_var   = tk.StringVar(value="PAUT")
        self.date_var      = tk.StringVar(value=datetime.date.today().strftime('%Y-%m-%d'))
        self.section_var   = tk.StringVar()
        self.line_var      = tk.StringVar()
        self.joint_var     = tk.StringVar()
        self.welder_var    = tk.StringVar()
        self.result_var    = tk.StringVar(value="합격")

        # ── 공정별 독립 관경·규격 변수 ──
        self.size_vars = {
            'PAUT': tk.StringVar(),
            'RT':   tk.StringVar(),
            'MT':   tk.StringVar(),
            'PT':   tk.StringVar(),
        }
        self.spec_vars = {
            'PAUT': tk.StringVar(),
            'RT':   tk.StringVar(),
            'MT':   tk.StringVar(),
            'PT':   tk.StringVar(),
        }

        # ── 공정별 전용 변수 ──
        self.shoot1_var    = tk.StringVar(value="-")
        self.shoot2_var    = tk.StringVar(value="-")
        self.shoot3_var    = tk.StringVar(value="-")
        self.shoot4_var    = tk.StringVar(value="-")
        self.rt_ori_var    = tk.StringVar()
        self.rt_re_var     = tk.StringVar()
        self.paut_ori_var  = tk.StringVar()
        self.paut_re_var   = tk.StringVar()
        self.mt_len_var    = tk.StringVar()
        self.pt_len_var    = tk.StringVar()
        self.remark_var    = tk.StringVar()

        # ── 고정(sticky) 변수 ──
        self.sticky_date    = tk.BooleanVar(value=True)
        self.sticky_section = tk.BooleanVar(value=True)
        self.sticky_line    = tk.BooleanVar(value=True)
        self.sticky_welder  = tk.BooleanVar(value=True)
        self.sticky_size    = tk.BooleanVar(value=True)
        self.sticky_spec    = tk.BooleanVar(value=True)

        # size_vars에 trace 연결 → PAUT/MT/PT 관경 선택 시 검사길이 자동 입력
        for proc_key in ('PAUT', 'MT', 'PT'):
            self.size_vars[proc_key].trace_add(
                'write',
                lambda *_, pk=proc_key: self._on_size_changed(pk)
            )

        self._build_ui()
        self._refresh_fields()

    def _on_size_changed(self, proc_key):
        """PAUT/MT/PT 관경 변경 시 검사길이 자동 입력"""
        size = self.size_vars[proc_key].get().strip()
        length = SIZE_LENGTH.get(size)
        if length is None:
            return
        length_str = str(length)
        if proc_key == 'PAUT':
            self.paut_ori_var.set(length_str)
        elif proc_key == 'MT':
            self.mt_len_var.set(length_str)
        elif proc_key == 'PT':
            self.pt_len_var.set(length_str)

    # ──────────────────────────────────────────────
    def _build_ui(self):
        # ① 공정 선택 바
        top = ttk.LabelFrame(self, text="① 공정 선택")
        top.pack(fill='x', padx=10, pady=5)
        for proc in ["PAUT", "RT", "MT", "PT"]:
            ttk.Radiobutton(top, text=proc, value=proc,
                            variable=self.process_var,
                            command=self._refresh_fields).pack(side='left', padx=15, pady=4)

        # ② 입력 폼
        self.form_lf = ttk.LabelFrame(self, text="② 데이터 입력  (☑ 고정 = 저장 후에도 값 유지)")
        self.form_lf.pack(fill='both', expand=False, padx=10, pady=5)

        # ③ 저장 버튼
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill='x', padx=10, pady=6)
        save_btn = ttk.Button(btn_frame,
                              text="💾  엑셀에 저장 및 다음 입력  (Enter)",
                              command=self._save_data)
        save_btn.pack(side='right', ipadx=20, ipady=8)
        self.bind('<Return>', lambda e: self._save_data())

        # ④ 미리보기
        prev_lf = ttk.LabelFrame(self, text="③ 최근 입력 기록")
        prev_lf.pack(fill='both', expand=True, padx=10, pady=5)
        cols = ("공정", "촬영일자", "Line No.", "Joint", "합부")
        self.tree = ttk.Treeview(prev_lf, columns=cols, show='headings', height=6)
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=120, anchor='center')
        sb = ttk.Scrollbar(prev_lf, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')

    # ──────────────────────────────────────────────
    def _row(self, parent, row, label, var,
             sticky_var=None, is_combo=False, values=None, width=28, readonly=False, is_date=False):
        ttk.Label(parent, text=label, anchor='e').grid(
            row=row, column=0, sticky='e', padx=(8,3), pady=4)
        if is_date and DateEntry:
            w = DateEntry(parent, textvariable=var, width=width-2,
                          date_pattern='yyyy-mm-dd', background='darkblue',
                          foreground='white', borderwidth=2)
        elif is_combo:
            w = ttk.Combobox(parent, textvariable=var, values=values or [], width=width)
            if readonly:
                w.configure(state='readonly')
        else:
            w = ttk.Entry(parent, textvariable=var, width=width)
        w.grid(row=row, column=1, sticky='w', padx=3, pady=4)
        if sticky_var is not None:
            ttk.Checkbutton(parent, text="고정", variable=sticky_var).grid(
                row=row, column=2, sticky='w', padx=5)
        return w

    def _pair_row(self, parent, row, label, var_ori, var_re):
        """ORI / RE' 두 칸을 한 행에"""
        ttk.Label(parent, text=label, anchor='e').grid(
            row=row, column=0, sticky='e', padx=(8,3), pady=4)
        sub = ttk.Frame(parent)
        sub.grid(row=row, column=1, sticky='w')
        ttk.Label(sub, text="ORI:").pack(side='left')
        ttk.Entry(sub, textvariable=var_ori, width=10).pack(side='left', padx=(2,8))
        ttk.Label(sub, text="RE':").pack(side='left')
        ttk.Entry(sub, textvariable=var_re, width=10).pack(side='left', padx=2)

    # ──────────────────────────────────────────────
    def _refresh_fields(self):
        # 기존 위젯 제거
        for w in self.form_lf.winfo_children():
            w.destroy()

        proc = self.process_var.get()
        f = self.form_lf

        r = 0
        # ── 공통 필드 ──
        # 공정 표시 (읽기전용 - 라디오버튼으로 선택)
        ttk.Label(f, text="공정:", anchor='e').grid(row=r, column=0, sticky='e', padx=(8,3), pady=4)
        ttk.Label(f, text=proc, font=('맑은 고딕', 11, 'bold'),
                  foreground='#1a6bbf').grid(row=r, column=1, sticky='w', padx=5)
        r += 1

        self._row(f, r, "촬영 일자:", self.date_var, self.sticky_date, is_date=True, width=14); r+=1
        self._row(f, r, "Section:", self.section_var, self.sticky_section); r+=1   # 전 공정 표시
        self._row(f, r, "Line No.:", self.line_var, self.sticky_line, width=40); r+=1
        self.joint_entry = self._row(f, r, "Joint No.:", self.joint_var); r+=1
        self._row(f, r, "용접사 번호:", self.welder_var, self.sticky_welder); r+=1
        # ★ 공정별 독립 관경/규격 변수 사용
        self._row(f, r, "관경 (Size):", self.size_vars[proc], self.sticky_size,
                  is_combo=True, values=SIZES, width=12, readonly=False); r+=1
        self._row(f, r, "규격 (주간/야간):", self.spec_vars[proc], self.sticky_spec,
                  is_combo=True, values=["주간", "야간", "주야간", "야간(주간)", "주간(야간)"],
                  width=14, readonly=False); r+=1
        self._row(f, r, "합부 판정:", self.result_var,
                  is_combo=True, values=["합격", "불합격"], width=12, readonly=True); r+=1

        # ── 공정별 전용 ──
        if proc == "RT":
            # 촬영구간 4개
            ttk.Label(f, text="촬영구간(1~4):", anchor='e').grid(
                row=r, column=0, sticky='e', padx=(8,3), pady=4)
            sub = ttk.Frame(f)
            sub.grid(row=r, column=1, sticky='w'); r+=1
            for i, var in enumerate([self.shoot1_var, self.shoot2_var,
                                      self.shoot3_var, self.shoot4_var], 1):
                ttk.Label(sub, text=f"{i}:").pack(side='left')
                ttk.Entry(sub, textvariable=var, width=6).pack(side='left', padx=(1,6))
            self._pair_row(f, r, "RT 촬영매수:", self.rt_ori_var, self.rt_re_var); r+=1

        elif proc in ["PAUT", "MT", "PT"]:
            # 촬영구간 1개 (통합)
            self._row(f, r, "촬영구간(통합):", self.shoot1_var, width=28); r+=1
            
            if proc == "PAUT":
                self._pair_row(f, r, "PAUT 길이(m):", self.paut_ori_var, self.paut_re_var); r+=1
            elif proc == "MT":
                self._row(f, r, "MT 길이(m):", self.mt_len_var); r+=1
            elif proc == "PT":
                self._row(f, r, "PT 길이(m):", self.pt_len_var); r+=1

        self._row(f, r, "비고:", self.remark_var, width=40); r+=1

        self.joint_entry.focus()

    # ──────────────────────────────────────────────
    def _save_data(self):
        proc = self.process_var.get()

        if not os.path.exists(self.excel_path):
            messagebox.showerror("오류", f"파일이 없습니다:\n{self.excel_path}")
            return

        try:
            wb = openpyxl.load_workbook(self.excel_path)
            if SHEET_NAME not in wb.sheetnames:
                messagebox.showerror("오류", f"'{SHEET_NAME}' 시트가 없습니다.")
                return
            ws = wb[SHEET_NAME]

            # ── 빈 행 탐색 (데이터 시작 행부터) ──
            insert_row = DATA_START_ROW
            while insert_row <= 5000:
                # 순번 또는 Line No. 가 비어있으면 빈 행
                if not ws.cell(row=insert_row, column=COL['순번']).value and \
                   not ws.cell(row=insert_row, column=COL['Line No.']).value:
                    break
                insert_row += 1

            # ── 순번 계산 ──
            seq_no = insert_row - DATA_START_ROW + 1

            # ── 셀 쓰기 ──
            def w(col_key, value):
                ws.cell(row=insert_row, column=COL[col_key]).value = value

            w('순번',       seq_no)
            w('촬영일자',   self.date_var.get())
            w('공정',       proc)
            w('Section',    self.section_var.get())
            w('Line No.',   self.line_var.get())
            w('Joint',      self.joint_var.get())
            w('용접사번호', self.welder_var.get())
            w('관경',       self.size_vars[proc].get())
            w('규격',       self.spec_vars[proc].get())
            w('합부',       self.result_var.get())
            w('비고',       self.remark_var.get())

            if proc == "RT":
                w('구간1',   self.shoot1_var.get())
                w('구간2',   self.shoot2_var.get())
                w('구간3',   self.shoot3_var.get())
                w('구간4',   self.shoot4_var.get())
                w('RT_ORI',  self.rt_ori_var.get())
                w("RT_RE'",  self.rt_re_var.get())
            elif proc == "PAUT":
                w('구간1',   self.shoot1_var.get())
                w('PAUT_ORI',  self.paut_ori_var.get())
                w("PAUT_RE'",  self.paut_re_var.get())
            elif proc == "MT":
                w('구간1',   self.shoot1_var.get())
                w('MT길이',  self.mt_len_var.get())
            elif proc == "PT":
                w('구간1',   self.shoot1_var.get())
                w('MT길이',  self.pt_len_var.get())  # PT길이 → MT길이 컬럼 공유(행이 겹치지 않음)

            # ── 서식 적용 (정렬 + 테두리) ──
            hair = Side(style='hair')
            thin = Side(style='thin')
            for c in range(1, MAX_COL + 1):
                cell = ws.cell(row=insert_row, column=c)
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)
                cell.font = Font(name='맑은 고딕', size=9)
                left_s  = thin if c == 1 else hair
                right_s = thin if c == MAX_COL else hair
                cell.border = Border(left=left_s, right=right_s, top=hair, bottom=hair)

            wb.save(self.excel_path)

            # ── 미리보기 갱신 ──
            self.tree.insert("", 0, values=(
                proc,
                self.date_var.get(),
                self.line_var.get(),
                self.joint_var.get(),
                self.result_var.get()
            ))

            # ── Joint 자동 증가 ──
            self._auto_increment_joint()

            # ── 비고정 필드 초기화 ──
            self._clear_fields()

        except PermissionError:
            messagebox.showerror("접근 거부",
                "엑셀 파일이 열려있습니다.\n닫은 후 다시 시도해주세요!")
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패:\n{e}")

    def _auto_increment_joint(self):
        j = self.joint_var.get()
        m = re.search(r'(\d+)$', j)
        if m:
            nxt = str(int(m.group(1)) + 1).zfill(len(m.group(1)))
            self.joint_var.set(j[:m.start()] + nxt)
        else:
            self.joint_var.set("")

    def _clear_fields(self):
        if not self.sticky_date.get():    self.date_var.set("")
        if not self.sticky_section.get(): self.section_var.set("")
        if not self.sticky_line.get():    self.line_var.set("")
        if not self.sticky_welder.get():  self.welder_var.set("")
        if not self.sticky_size.get():
            for v in self.size_vars.values(): v.set("")
        if not self.sticky_spec.get():
            for v in self.spec_vars.values(): v.set("")
        # 전용 필드 초기화
        self.rt_ori_var.set("")
        self.rt_re_var.set("")
        self.paut_ori_var.set("")
        self.paut_re_var.set("")
        self.mt_len_var.set("")
        self.pt_len_var.set("")
        self.remark_var.set("")
        # 촬영구간 리셋
        for v in [self.shoot1_var, self.shoot2_var,
                  self.shoot3_var, self.shoot4_var]:
            v.set("-")


def open_data_entry_ui(parent):
    DataEntryUI(parent)
