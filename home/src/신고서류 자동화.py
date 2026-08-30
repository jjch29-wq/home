import sys
import os
import glob
import openpyxl
import math
from openpyxl.styles import Font, PatternFill, Alignment
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                               QHBoxLayout, QPushButton, QLabel, QFileDialog, 
                               QTextEdit, QMessageBox, QProgressBar)
from PySide6.QtCore import Qt, QThread, Signal
import win32com.client as win32
import diagram_generator
import math_generator
import json
import re
from datetime import datetime

# 헤더(치환할 태그들) 리스트
HEADERS = [
    "폴더명", "발주처", "공사명", "공사명_상세", "현장주소", "주변특징", "민가거리", 
    "주야간작업", "관공서거리", "경찰서이름", "소방서이름", "발주처담당자", "담당자이름", 
    "담당자부서", "담당자휴대폰", "담당자사무실전화", "담당자이메일", "사업자번호", 
    "안전관리자", "사용선원", "현장전화번호", "작업조", "시작일", "종료일", 
    "운반경로", "총이동거리", "소요시간", 
    "작업장구분", "작업장크기", "차폐물", 
    "거리_상단", "거리_하단", "거리_좌측", "거리_우측", "배관길이", "배관폭", "매설깊이", "관경",
    "재질", "관경범위", "최대두께", "총작업량", 
    "조사대상_OUT", "조사방법", "조사방향", "Stop시간", "Shooting시간",
    "계산식_선원종류", "계산식_방사능", "계산식_콜리메이터_두께", "계산식_콜리메이터_반가층",
    "계산식_납패드_두께", "계산식_납패드_반가층", "계산식_토양_반가층",
    "지도이미지폴더"
]

class HWPGeneratorThread(QThread):
    log_signal = Signal(str)
    progress_signal = Signal(int)
    finished_signal = Signal(bool)

    def __init__(self, template_files, excel_path, output_base_dir, global_img_folder=""):
        super().__init__()
        self.template_files = template_files
        self.excel_path = excel_path
        self.output_base_dir = output_base_dir
        self.global_img_folder = global_img_folder

    def run(self):
        try:
            self.log_signal.emit("엑셀 데이터를 읽는 중...")
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            ws = wb.active
            
            # 첫 번째 행은 헤더
            headers = [cell.value for cell in ws[1]]
            
            # 두 번째 행부터 데이터 처리
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            total_rows = len([row for row in data_rows if row[0] is not None])
            
            if total_rows == 0:
                self.log_signal.emit("엑셀에 작성된 데이터가 없습니다.")
                self.finished_signal.emit(False)
                return

            self.log_signal.emit(f"총 {total_rows}개의 현장 서류를 생성합니다.")
            
            hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
            hwp.RegisterModule("FilePathCheckDLL", "SecurityModule")
            
            if not self.template_files:
                self.log_signal.emit("선택된 템플릿 파일이 없습니다!")
                self.finished_signal.emit(False)
                return

            current_progress = 0
            
            def parse_num(val_str, default_val):
                if not val_str: return default_val
                nums = re.findall(r'[\d.]+', str(val_str))
                if nums: return float(nums[0])
                return default_val
                
            def parse_date(date_str):
                date_str = str(date_str).strip()
                try:
                    if "-" in date_str: return datetime.strptime(date_str, "%Y-%m-%d")
                    else: return datetime.strptime(date_str, "%Y.%m.%d")
                except: return None
            
            warnings_list = []
            
            for row_idx, row in enumerate(data_rows):
                if row[0] is None:
                    continue # 폴더명이 없으면 스킵
                
                # 딕셔너리로 변환
                row_data = {headers[i]: (str(row[i]) if row[i] is not None else "") for i in range(len(headers))}
                folder_name = row_data.get("폴더명", f"새로운현장_{row_idx}")
                
                # 날짜 자동 계산
                start_date = parse_date(row_data.get("시작일", ""))
                end_date = parse_date(row_data.get("종료일", ""))
                
                if start_date and end_date:
                    delta = end_date - start_date
                    months = delta.days // 30
                    years = months // 12
                    rem_months = months % 12
                    
                    period_str = ""
                    if years > 0: period_str += f"{years}년 "
                    if rem_months > 0: period_str += f"{rem_months}개월"
                    elif years == 0: period_str = f"{delta.days}일"
                        
                    row_data["작업기간"] = period_str.strip()
                    row_data["시작일_단축"] = start_date.strftime("%y.%m.%d")
                    row_data["종료일_단축"] = end_date.strftime("%y.%m.%d")
                else:
                    row_data["작업기간"] = row_data.get("작업기간", "")
                    row_data["시작일_단축"] = row_data.get("시작일", "")
                    row_data["종료일_단축"] = row_data.get("종료일", "")
                
                try:
                    pipe_dia_val = parse_num(row_data.get("관경", "150"), 150)
                    shield_size = int(pipe_dia_val * 3.14 * (2/3))
                    row_data["납패드계산"] = f"- {int(pipe_dia_val)}mm이하 (관경) x 3.14 x 2/3 = {shield_size}mm"
                except Exception:
                    row_data["납패드계산"] = "- 150mm이하 (관경) x 3.14 x 2/3 = 314mm"
                    
                try:
                    workspace_size = str(row_data.get("작업장크기", "5 x 14")).strip()
                    nums = re.findall(r'[\d.]+', workspace_size)
                    if len(nums) >= 2:
                        w1, w2 = nums[0], nums[1]
                    else:
                        w1, w2 = "5", "14"
                    row_data["통제구역계산"] = f"- 방사선 관리·감시 구역은 {w1}m x {w2}m 이나 매설구간으로부터 종방향 최소 {w1}m 이상, 횡방향 {w2}m 이상에서 이동식 펜스를 설치하여 일반인의 접근을 통제함."
                except Exception:
                    w1, w2 = "5", "14"
                    row_data["통제구역계산"] = "- 방사선 관리·감시 구역은 5m x 14m 이나 매설구간으로부터 종방향 최소 5m 이상, 횡방향 14m 이상에서 이동식 펜스를 설치하여 일반인의 접근을 통제함."
                try:
                    w1_half = float(w1) / 2
                    w2_half = float(w2) / 2
                    row_data["횡축거리"] = f"{w1_half:g}"
                    row_data["종축거리"] = f"{w2_half:g}"
                except:
                    row_data["횡축거리"] = "2.5"
                    row_data["종축거리"] = "7"
                
                try:
                    trench_width = parse_num(row_data.get("배관폭", "1000"), 1000)
                    W = trench_width / 2
                    right_dist = parse_num(row_data.get("거리_우측", "2000"), 2000)
                    person_x_math = W + right_dist
                    pipe_depth = parse_num(row_data.get("매설깊이", "1500"), 1500)
                        
                    person_height = 2000
                    slope = (pipe_depth + person_height) / person_x_math if person_x_math > 0 else 1.4
                    angle_rad = math.atan(slope)
                    A_math = slope * W
                    B_math = pipe_depth - A_math
                    C_math = B_math / math.sin(angle_rad) if angle_rad > 0 else 0
                    row_data["자동계산토양두께"] = f"{int(C_math)}mm"
                    row_data["계산식_토양_두께"] = f"{int(C_math)}"
                except Exception:
                    row_data["자동계산토양두께"] = "983mm"
                    row_data["계산식_토양_두께"] = "983"

                try:
                    total_shots = parse_num(row_data.get("총작업량", "2000"), 2000)
                    stop_time = parse_num(row_data.get("Stop시간", "35"), 35)
                    shoot_time = parse_num(row_data.get("Shooting시간", "4"), 4)
                    
                    stop_hr = (total_shots * stop_time) / 3600
                    shoot_hr = (total_shots * shoot_time) / 3600
                    
                    row_data["계산_Stop시간"] = f"{round(stop_hr, 1)}"
                    row_data["계산_Shooting시간"] = f"{round(shoot_hr, 2)}"
                except Exception:
                    row_data["계산_Stop시간"] = "19.4"
                    row_data["계산_Shooting시간"] = "2.22"

                out_val = row_data.get("조사대상_OUT", "O")
                row_data["체크_OUT"] = "☑" if out_val.strip().upper() == "O" else "□"
                
                method = row_data.get("조사방법", "이중벽 단상").strip()
                row_data["체크_이중벽단상"] = "☑이중벽 단상" if method == "이중벽 단상" else "□이중벽 단상"
                
                directions = row_data.get("조사방향", "하향, 상향, 측하향, 측방향, 측상향")
                row_data["체크_하향"] = "☑ 하향" if "하향" in directions else "□ 하향"
                row_data["체크_상향"] = "☑ 상향" if "상향" in directions else "□ 상향"
                row_data["체크_측하향"] = "☑ 측하향" if "측하향" in directions else "□ 측하향"
                row_data["체크_측방향"] = "☑ 측방향" if "측방향" in directions else "□ 측방향"
                row_data["체크_측상향"] = "☑ 측상향" if "측상향" in directions else "□ 측상향"
                
                daynight = row_data.get("주야간작업", "야간")
                if "주" in daynight:
                    row_data["체크_주"] = "☑주"
                    row_data["체크_야"] = "□야"
                else:
                    row_data["체크_주"] = "□주"
                    row_data["체크_야"] = "☑야"

                # 폴더를 새로 만들지 않고, 바탕화면(또는 선택한 저장경로)에 바로 저장
                target_dir = self.output_base_dir
                file_prefix = folder_name
                
                self.log_signal.emit(f"\n[{folder_name}] 서류 생성 시작...")
                
                for template_file in self.template_files:
                    base_name = os.path.basename(template_file)
                    # 파일명 앞에 현장 이름(폴더명 컬럼값)을 붙여서 덮어쓰기 방지
                    target_file = os.path.join(target_dir, f"{file_prefix}_{base_name}")
                    
                    self.log_signal.emit(f"  - {base_name} 치환 중...")
                    
                    hwp.Open(template_file, "HWP", "forceopen:true")
                    
                    for key, val in row_data.items():
                        if key == "폴더명":
                            continue
                            
                        # {{키}} 형태로 찾아 바꾸기
                        find_str = f"{{{{{key}}}}}" 
                        
                        hwp.HAction.Run("MoveDocBegin") # 문서 맨 앞으로 커서 이동
                        hwp.HAction.GetDefault("AllReplace", hwp.HParameterSet.HFindReplace.HSet)
                        hwp.HParameterSet.HFindReplace.FindString = find_str
                        hwp.HParameterSet.HFindReplace.ReplaceString = val
                        hwp.HParameterSet.HFindReplace.IgnoreMessage = 1
                        hwp.HAction.Execute("AllReplace", hwp.HParameterSet.HFindReplace.HSet)
                        
                    # 1. 자동 생성 도면 (구역도) 그리기
                    from diagram_generator import create_radiation_diagram, create_side_diagram
                    diagram_path = ""
                    side_diagram_path = ""
                    try:
                        d_top = parse_num(row_data.get("거리_상단", 2000), 2000)
                        d_bot = parse_num(row_data.get("거리_하단", 2000), 2000)
                        d_left = parse_num(row_data.get("거리_좌측", 2000), 2000)
                        d_right = parse_num(row_data.get("거리_우측", 2000), 2000)
                        p_len = parse_num(row_data.get("배관길이", 10000), 10000)
                        p_wid = parse_num(row_data.get("배관폭", 1000), 1000)
                        p_dia = parse_num(row_data.get("관경", "150"), 150)
                        p_depth = parse_num(row_data.get("매설깊이", "1500"), 1500)
                            
                        diagram_path = create_radiation_diagram(d_top, d_bot, d_left, d_right, p_len, p_wid, os.path.join(target_dir, f"{file_prefix}_temp_diagram.jpg"))
                        side_diagram_path = create_side_diagram(d_left, d_right, p_wid, p_dia, p_depth, os.path.join(target_dir, f"{file_prefix}_side_diagram.jpg"))
                    except Exception as e:
                        self.log_signal.emit(f"    - 도면 생성 실패: {str(e)}")
                        diagram_path = ""

                    # 1.5 계산식(수식) 이미지 자동 생성
                    math_imgs_dir = os.path.join(target_dir, f"{file_prefix}_math_imgs")
                    math_paths = {}
                    try:
                        source_type = str(row_data.get("계산식_선원종류", "Se-75"))
                        activity = float(row_data.get("계산식_방사능", "60") or "60")
                        col_t = float(row_data.get("계산식_콜리메이터_두께", "11") or "11")
                        col_h = float(row_data.get("계산식_콜리메이터_반가층", "0.8") or "0.8")
                        pb_t = float(row_data.get("계산식_납패드_두께", "12") or "12")
                        pb_h = float(row_data.get("계산식_납패드_반가층", "1") or "1")
                        soil_t = float(row_data.get("계산식_토양_두께", "983") or "983")
                        soil_h = float(row_data.get("계산식_토양_반가층", "45") or "45")
                        d_ab = float(w1_half)
                        d_cd = float(w2_half)
                        
                        stop_hr = float(row_data.get("계산_Stop시간", "19.4"))
                        shoot_hr = float(row_data.get("계산_Shooting시간", "2.22"))
                        scatter_base = float(row_data.get("산란방사선_기본값", "27.1"))
                        
                        generated_math_paths, satisfaction = math_generator.generate_math_images(
                            source_type, activity, col_t, col_h, pb_t, pb_h, soil_t, soil_h, 
                            d_ab, d_cd, stop_hr, shoot_hr, math_imgs_dir, scatter_base
                        )
                        # 만족여부 텍스트 치환 (이미지보다 먼저 수행)
                        for key, val in satisfaction.items():
                            if val == "불만족" and key.startswith("만족여부"):
                                warnings_list.append(folder_name)
                                self.log_signal.emit(f"<font color='red'><b>[경고] {folder_name} 현장 선량률 기준치 초과 (불만족)</b></font>")
                            find_str = f"{{{{{key}}}}}"
                            hwp.HAction.Run("MoveDocBegin")
                            hwp.HAction.GetDefault("AllReplace", hwp.HParameterSet.HFindReplace.HSet)
                            hwp.HParameterSet.HFindReplace.FindString = find_str
                            hwp.HParameterSet.HFindReplace.ReplaceString = val
                            hwp.HParameterSet.HFindReplace.IgnoreMessage = 1
                            hwp.HAction.Execute("AllReplace", hwp.HParameterSet.HFindReplace.HSet)

                        for path in generated_math_paths:
                            if "diagram1" in path: math_paths["평가수식1"] = path
                            elif "diagram2" in path: math_paths["평가수식2"] = path
                            elif "diagram3" in path: math_paths["평가수식3"] = path
                            elif "equation1" in path: math_paths["계산식1"] = path
                            elif "equation2" in path: math_paths["계산식2"] = path
                            elif "equation3" in path: math_paths["계산식3"] = path
                            elif "eval_eq1" in path: math_paths["평가결과1"] = path
                            elif "eval_eq2" in path: math_paths["평가결과2"] = path
                            elif "max_dose" in path: math_paths["최대선량률"] = path
                            elif "dose_ab" in path: math_paths["피폭선량결과"] = path
                            elif "dose_cd" in path: math_paths["산란방사선결과"] = path
                            elif "eq_shooting" in path: math_paths["평가수식_Shooting"] = path
                            elif "eq_stop" in path: math_paths["평가수식_Stop"] = path
                            elif "eq_scatter" in path: math_paths["평가수식_Scatter"] = path
                            elif "eq_total" in path: math_paths["평가수식_Total"] = path
                            elif "total_result" in path: math_paths["최종피폭선량"] = path
                            
                    except Exception as e:
                        self.log_signal.emit(f"    - 수식 생성 실패: {str(e)}")

                    # 2. 이미지 삽입 로직
                    # 우선 GUI에서 선택한 폴더가 있으면 그것을 사용, 없으면 엑셀의 데이터 사용
                    img_folder = self.global_img_folder if self.global_img_folder else row_data.get("지도이미지폴더", "")
                    
                    # 삽입할 이미지 태그 목록
                    image_tags = [
                        "지도1", "지도2", "지도3", 
                        "부지사진1", "부지사진2",
                        "상세사진1", "상세사진2",
                        "구역도", "횡방향측면도", 
                        "계산식1", "계산식2", "계산식3", "계산식4",
                        "계산식5", "계산식6", "계산식7", "계산식8",
                        "평가결과1", "평가결과2", "평가수식1", "평가수식2",
                        "최대선량률", "피폭선량결과", "산란방사선결과",
                        "평가수식_Shooting", "평가수식_Stop",
                        "평가수식_Scatter", "평가수식_Total", "최종피폭선량"
                    ]
                    
                    for tag in image_tags:
                        placeholder = f"{{{{{tag}}}}}"
                        img_path = ""
                        
                        # 구역도 태그이고 자동 생성된 파일이 있다면 우선 사용
                        if tag == "구역도" and diagram_path and os.path.exists(diagram_path):
                            img_path = diagram_path
                        elif tag == "횡방향측면도" and side_diagram_path and os.path.exists(side_diagram_path):
                            img_path = side_diagram_path
                        elif (tag.startswith("계산식") or tag.startswith("평가결과") or tag.startswith("평가수식") or tag in ["최대선량률", "최종피폭선량"]) and tag in math_paths and os.path.exists(math_paths[tag]):
                            img_path = math_paths[tag]
                        # 그 외엔 이미지 폴더에서 찾기
                        elif img_folder and os.path.exists(img_folder):
                            for ext in [".jpg", ".png", ".jpeg"]:
                                temp_path = os.path.join(img_folder, f"{tag}{ext}")
                                if os.path.exists(temp_path):
                                    img_path = temp_path
                                    break
                        
                        if img_path:
                                hwp.HAction.Run("MoveDocBegin")
                                hwp.HAction.GetDefault("RepeatFind", hwp.HParameterSet.HFindReplace.HSet)
                                hwp.HParameterSet.HFindReplace.FindString = placeholder
                                hwp.HParameterSet.HFindReplace.IgnoreMessage = 1
                                # 태그를 찾았다면
                                if hwp.HAction.Execute("RepeatFind", hwp.HParameterSet.HFindReplace.HSet):
                                    hwp.HAction.Run("Delete") # 태그 지우기
                                    if tag == "구역도":
                                        hwp.InsertPicture(img_path, True, 1, False, False, 0, 150, 98)
                                    elif tag == "횡방향측면도":
                                        hwp.InsertPicture(img_path, True, 1, False, False, 0, 150, 105)
                                    elif tag == "지도1":
                                        hwp.InsertPicture(img_path, True, 1, False, False, 0, 150, 80)
                                    elif tag in ["지도2", "지도3"] or tag.startswith("부지사진") or tag.startswith("상세사진"):
                                        hwp.InsertPicture(img_path, True, 1, False, False, 0, 75, 55)
                                    elif tag.startswith("평가수식"):
                                        hwp.InsertPicture(img_path, True, 1, False, False, 0, 150, 10)
                                    elif tag.startswith("계산식"):
                                        hwp.InsertPicture(img_path, True, 1, False, False, 0, 150, 30)
                                    elif tag.startswith("평가결과"):
                                        # 평가결과는 작은 크기로 삽입
                                        hwp.InsertPicture(img_path, True, 1, False, False, 0, 45, 15)
                                    elif tag in ["최대선량률", "피폭선량결과", "산란방사선결과", "최종피폭선량"]:
                                        # 최대선량률 등은 줄글 중간이나 좁은 칸에 들어가므로 아주 작게 삽입
                                        hwp.InsertPicture(img_path, True, 1, False, False, 0, 30, 8)
                                    else:
                                        # 지도/부지사진 등은 지정된 15cm x 10cm 크기로 고정 (크기 옵션 1, 150mm x 100mm)
                                        hwp.InsertPicture(img_path, True, 1, False, False, 0, 150, 100)
                                        
                                    self.log_signal.emit(f"    - {placeholder} ➡️ {os.path.basename(img_path)} 삽입 완료")
                                    
                    hwp.SaveAs(target_file, "HWP")
                    hwp.Clear(1)
                
                current_progress += 1
                self.progress_signal.emit(int(current_progress / total_rows * 100))
                self.log_signal.emit(f"[{folder_name}] 생성 완료!")

            hwp.Quit()
            if warnings_list:
                self.log_signal.emit(f"\n<font color='red'><b>총 {len(warnings_list)}건의 현장에서 방사선량 '불만족'이 발생했습니다. 설계를 확인하세요!</b></font>")
            self.log_signal.emit("\n모든 작업이 성공적으로 완료되었습니다!")
            self.finished_signal.emit(True)
            
        except Exception as e:
            self.log_signal.emit(f"\n에러 발생: {str(e)}")
            self.finished_signal.emit(False)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("동위원소 개설 신고 서류 자동화")
        self.resize(700, 500)
        
        self.template_files = []
        self.excel_path = ""
        self.img_folder = ""
        self.last_output_dir = ""
        self.settings_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "settings.json")
        
        self.initUI()
        self.load_settings()
        
    def initUI(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        
        # 1. 엑셀 DB 템플릿 생성 버튼
        btn_layout1 = QHBoxLayout()
        self.btn_create_excel = QPushButton("1. 엑셀 입력 폼(양식) 만들기")
        self.btn_create_excel.setMinimumHeight(40)
        self.btn_create_excel.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; border-radius: 5px;")
        self.btn_create_excel.clicked.connect(self.create_excel_template)
        btn_layout1.addWidget(self.btn_create_excel)
        layout.addLayout(btn_layout1)
        
        # 2. 템플릿 및 엑셀 선택
        layout.addWidget(QLabel("\n2. 파일 선택"))
        
        path_layout1 = QHBoxLayout()
        self.btn_sel_template = QPushButton("HWP 템플릿 파일 선택 (여러개 가능)")
        self.btn_sel_template.clicked.connect(self.select_template_files)
        self.lbl_template = QLabel("선택 안됨")
        path_layout1.addWidget(self.btn_sel_template)
        path_layout1.addWidget(self.lbl_template, 1)
        layout.addLayout(path_layout1)
        
        path_layout2 = QHBoxLayout()
        self.btn_sel_excel = QPushButton("작성된 엑셀 파일 선택")
        self.btn_sel_excel.clicked.connect(self.select_excel_file)
        self.lbl_excel = QLabel("선택 안됨")
        path_layout2.addWidget(self.btn_sel_excel)
        path_layout2.addWidget(self.lbl_excel, 1)
        layout.addLayout(path_layout2)
        
        path_layout3 = QHBoxLayout()
        self.btn_sel_img = QPushButton("지도 사진 폴더 선택 (선택사항)")
        self.btn_sel_img.setStyleSheet("background-color: #ff9800; color: white; font-weight: bold;")
        self.btn_sel_img.clicked.connect(self.select_img_folder)
        self.lbl_img = QLabel("선택 안됨 (엑셀 내용 우선)")
        path_layout3.addWidget(self.btn_sel_img)
        path_layout3.addWidget(self.lbl_img, 1)
        layout.addLayout(path_layout3)
        
        # 3. 자동 생성 실행
        layout.addWidget(QLabel("\n3. 실행"))
        self.btn_run = QPushButton("신고 서류 일괄 자동 생성")
        self.btn_run.setMinimumHeight(50)
        self.btn_run.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold; font-size: 14px; border-radius: 5px;")
        self.btn_run.clicked.connect(self.run_generation)
        layout.addWidget(self.btn_run)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)
        
        # 4. 로그
        layout.addWidget(QLabel("진행 상황 로그"))
        self.log_view = QTextEdit()
        self.log_view.setReadOnly(True)
        layout.addWidget(self.log_view)
        
    def log(self, msg):
        self.log_view.append(msg)
        
    def create_excel_template(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "엑셀 양식 저장", "신고서작성_DB.xlsx", "Excel Files (*.xlsx)")
        if file_path:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "신고데이터"
            
            # 헤더 스타일
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_font = Font(bold=True)
            align = Alignment(horizontal="center", vertical="center")
            
            for col_num, header in enumerate(HEADERS, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
                
            # 샘플 데이터 1줄 추가
            sample_data = [
                "25.04.03_롯데건설(주)", "롯데건설(주)", "롯데바이오로직스 송도", "롯데바이오로직스 Project 송도1공장", 
                "인천광역시 송도동 418", "신 매립지", "800m", "야간 작업", "10km", "인천 송도 경찰서", "인천 송도 소방서", 
                "고준호(010-5090-8871)", "고준호", "롯데건설 품질팀", "010-5090-8871", "032-822-0988", 
                "junhoko@lotte.net", "114-81-16377", "김춘호(010-7763-3436)", "Se-75 60Ci x 1EA", 
                "032-549-8457", "3인 1조 (1개조)", "2025.04.03", "2025.05.07",
                "소래로 → 아암대로 → 인천신항대로", "15km", "30분", 
                "야외(매설배관)", "5 x 14", "황축토양(983)",
                "2000", "2000", "2000", "2000", "10000", "1000", "1500", "150",
                "C/S", "2~10''", "~5.5", "2000", 
                "O", "이중벽 단상", "하향, 상향, 측하향, 측방향, 측상향", "35", "4",
                "Se-75", "60", "11", "0.8", "12", "1", "45",
                r"C:\Users\jjch2\Desktop\신고서_마스터템플릿\지도사진들"
            ]
            for col_num, val in enumerate(sample_data, 1):
                ws.cell(row=2, column=col_num).value = val
                
            wb.save(file_path)
            self.log(f"엑셀 양식이 저장되었습니다: {file_path}")
            self.log("엑셀 파일을 열어서 현장 정보를 입력해 주세요!\n")
            
    def select_template_files(self):
        start_dir = os.path.dirname(self.template_files[0]) if self.template_files else ""
        files, _ = QFileDialog.getOpenFileNames(self, "마스터 템플릿 파일 선택", start_dir, "HWP Files (*.hwp)")
        if files:
            self.template_files = files
            self.lbl_template.setText(f"{len(files)}개 파일 선택됨")
            self.log(f"템플릿 파일 {len(files)}개 선택됨")
            self.save_settings()
            
    def select_excel_file(self):
        start_dir = os.path.dirname(self.excel_path) if self.excel_path else ""
        file_path, _ = QFileDialog.getOpenFileName(self, "작성된 엑셀 파일 선택", start_dir, "Excel Files (*.xlsx *.xls)")
        if file_path:
            self.excel_path = file_path
            self.lbl_excel.setText(file_path)
            self.log(f"엑셀 파일 선택됨: {file_path}")
            self.save_settings()
            
    def select_img_folder(self):
        start_dir = self.img_folder if self.img_folder else ""
        dir_path = QFileDialog.getExistingDirectory(self, "지도 사진(지도1, 지도2, 지도3)이 있는 폴더 선택", start_dir)
        if dir_path:
            self.img_folder = dir_path
            self.lbl_img.setText(dir_path)
            self.log(f"지도 사진 폴더 선택됨: {dir_path}")
            self.save_settings()
            
    def run_generation(self):
        if not self.template_files or not self.excel_path:
            QMessageBox.warning(self, "경고", "템플릿 파일과 엑셀 파일을 모두 선택해 주세요.")
            return
            
        output_base_dir = QFileDialog.getExistingDirectory(self, "완성된 서류를 저장할 최상위 폴더 선택", self.last_output_dir)
        if not output_base_dir:
            return
            
        self.last_output_dir = output_base_dir
        self.save_settings()
            
        self.btn_run.setEnabled(False)
        self.progress_bar.setValue(0)
        self.log_view.clear()
        
        self.thread = HWPGeneratorThread(self.template_files, self.excel_path, output_base_dir, self.img_folder)
        self.thread.log_signal.connect(self.log)
        self.thread.progress_signal.connect(self.progress_bar.setValue)
        self.thread.finished_signal.connect(self.on_finished)
        self.thread.start()
        
    def on_finished(self, success):
        self.btn_run.setEnabled(True)
        if success:
            QMessageBox.information(self, "완료", "모든 서류 생성이 완료되었습니다!")
        else:
            QMessageBox.critical(self, "오류", "작업 중 오류가 발생했습니다. 로그를 확인하세요.")
            
    def load_settings(self):
        if os.path.exists(self.settings_file):
            try:
                with open(self.settings_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    
                if "template_files" in data and data["template_files"]:
                    self.template_files = data["template_files"]
                    self.lbl_template.setText(f"{len(self.template_files)}개 파일 선택됨")
                    
                if "excel_path" in data and os.path.exists(data["excel_path"]):
                    self.excel_path = data["excel_path"]
                    self.lbl_excel.setText(self.excel_path)
                    
                if "img_folder" in data and os.path.exists(data["img_folder"]):
                    self.img_folder = data["img_folder"]
                    self.lbl_img.setText(self.img_folder)
                    
                if "last_output_dir" in data:
                    self.last_output_dir = data["last_output_dir"]
            except Exception as e:
                print(f"설정 불러오기 실패: {e}")
                
    def save_settings(self):
        data = {
            "template_files": self.template_files,
            "excel_path": self.excel_path,
            "img_folder": self.img_folder,
            "last_output_dir": self.last_output_dir
        }
        try:
            with open(self.settings_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"설정 저장 실패: {e}")
            
    def closeEvent(self, event):
        self.save_settings()
        super().closeEvent(event)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
