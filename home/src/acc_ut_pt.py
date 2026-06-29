import os
import openpyxl

desktop = r'C:\Users\-\OneDrive\바탕 화면'
files = [f for f in os.listdir(desktop) if '산출내역서' in f and f.endswith('.xlsx') and not f.startswith('~$')]

if files:
    filepath = os.path.join(desktop, files[0])
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    sums = {
        "UT": {"수송배관(주배관)": {"일반": 0.0, "야간": 0.0, "휴일": 0.0},
               "플랜트(관리소)": {"일반": 0.0, "야간": 0.0, "휴일": 0.0}},
        "PT": {"수송배관(주배관)": {"일반": 0.0, "야간": 0.0, "휴일": 0.0},
               "플랜트(관리소)": {"일반": 0.0, "야간": 0.0, "휴일": 0.0}}
    }
    
    for sheet_name in wb.sheetnames:
        if 'UT' in sheet_name or 'PT' in sheet_name:
            ws = wb[sheet_name]
            mode = "UT" if "UT" in sheet_name else "PT"
            current_shift = "일반"
            
            for r in range(1, ws.max_row + 1):
                row_str = " ".join([str(ws.cell(row=r, column=c).value) for c in range(1, 20) if ws.cell(row=r, column=c).value is not None])
                if '주간' in row_str: current_shift = "일반"
                elif '야간' in row_str: current_shift = "야간"
                elif '휴일' in row_str: current_shift = "휴일"
                
                val_col5 = str(ws.cell(row=r, column=5).value) if ws.cell(row=r, column=5).value else ""
                cat = "수송배관(주배관)" if '주배관' in val_col5 else ("플랜트(관리소)" if '관리소' in val_col5 else None)
                
                if cat:
                    val = 0
                    if mode == "PT":
                        # PT total is usually in column 18
                        val = ws.cell(row=r, column=18).value
                        if val is None: val = ws.cell(row=r, column=16).value
                    else:
                        # UT total might be in column 16, 14, or 18.
                        # Let's check col 18, then 16, then 14
                        val = ws.cell(row=r, column=18).value
                        if val is None: val = ws.cell(row=r, column=16).value
                        if val is None: val = ws.cell(row=r, column=14).value
                    
                    try:
                        sums[mode][cat][current_shift] += float(val)
                    except (ValueError, TypeError):
                        pass
                        
    for m in ["UT", "PT"]:
        print(f"\n--- {m} ---")
        for cat in sums[m]:
            for shift in sums[m][cat]:
                print(f"{cat} {shift}: {round(sums[m][cat][shift], 2)}")
