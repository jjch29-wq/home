import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from copy import copy
import sys

file_path = r"C:\Users\jjch2\Desktop\4.4.1_위험성평가표(RT_표준양식).xlsx"
save_path = r"C:\Users\jjch2\Desktop\4.4.1_위험성평가표(RT_표준양식)_업데이트.xlsx"

try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    new_data = [
        ("비파괴검사", "물리적 요인", "가이드튜브 꺾임, 장비 결함 등으로 동위원소 선원 미회수 시 고선량 피폭 위험", "사용 전 장비 점검 / 비상용 차폐복 및 집게(Handling tong) 비치 / 비상훈련", 1, 3, "III"),
        ("비파괴검사", "작업환경 요인", "야간 검사 시 조명 부족 및 시야 미확보로 인한 전도, 추락, 장비 충돌 위험", "작업구간 투광기 설치 / 안전조끼(반사띠) 착용 / 개인용 랜턴 지급", 2, 2, "IV"),
        ("사전준비", "화학(물질)적 요인", "배관 내부 및 깊은 트렌치 내부 출입 시 산소 결핍 또는 유해가스 체류로 인한 질식", "출입 전 산소/유해가스 농도 측정 / 가스마스크 등 개인보호구 지급", 2, 3, "V"),
        ("비파괴검사", "전기적 요인", "야간 조명등 및 검사장비 케이블 피복 손상, 누전, 우천 시 감전 위험", "누전차단기 부착 분전반 사용 / 전선 바닥 방치 금지(거치) / 우천 시 작업 통제", 2, 2, "IV"),
        ("이동", "기계적(설비)적 요인", "야간 암실 차량 주·정차 및 이동 시 시야 확보 불량으로 근로자 및 타 장비 충돌", "신호수(유도자) 배치 / 차량 경광등 작동 / 지정된 안전 통제구역 주차", 2, 3, "V"),
        ("작업 후 정리", "물리적 요인", "차량 이동 중 동위원소 저장용기 전도/낙하로 인한 방사능 유출 및 장비 파손", "전용 운반차량 사용(경고표지 부착) / 저장소 시건장치 및 고정 상태 확인", 1, 3, "III")
    ]

    # 마지막으로 데이터가 있는 행 찾기 (C열 기준)
    last_row = 10
    for row in range(10, ws.max_row + 20):
        cell_val = ws.cell(row=row, column=3).value
        if cell_val is not None and str(cell_val).strip() != "":
            last_row = row

    start_row = last_row + 1

    for i, data in enumerate(new_data):
        row_idx = start_row + i
        
        # 값 입력
        ws.cell(row=row_idx, column=1).value = data[0]  # A (세부작업)
        ws.cell(row=row_idx, column=2).value = data[1]  # B (위험분류)
        ws.cell(row=row_idx, column=3).value = data[2]  # C (유해위험요인)
        ws.cell(row=row_idx, column=7).value = data[3]  # G (현재안전보건조치)
        ws.cell(row=row_idx, column=10).value = data[4] # J (빈도)
        ws.cell(row=row_idx, column=11).value = data[5] # K (강도)
        ws.cell(row=row_idx, column=12).value = data[6] # L (등급)
        
        # 셀 병합 (C~F, G~I)
        ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=6)
        ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=9)
        ws.merge_cells(start_row=row_idx, start_column=12, end_row=row_idx, end_column=14)
        ws.merge_cells(start_row=row_idx, start_column=17, end_row=row_idx, end_column=20)
        ws.merge_cells(start_row=row_idx, start_column=21, end_row=row_idx, end_column=24)
        
        # 스타일 복사 (이전 행에서 복사)
        for col in range(1, 25):
            source_cell = ws.cell(row=last_row, column=col)
            target_cell = ws.cell(row=row_idx, column=col)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

    wb.save(save_path)
    print("Success: " + save_path)

except Exception as e:
    print(f"Error: {e}")
    sys.exit(1)
