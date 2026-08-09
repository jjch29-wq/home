import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class NDTSummaryExporter:
    def __init__(self, history):
        self.history = history
        self.font_normal = Font(name='맑은 고딕', size=10)
        self.font_bold = Font(name='맑은 고딕', size=10, bold=True)
        
        self.align_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
        
        thin = Side(border_style="thin", color="000000")
        self.border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.fill_header = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    def generate(self, output_path):
        wb = openpyxl.Workbook()
        
        # Group data by method
        grouped_data = {}
        
        # Sort dates
        sorted_dates = sorted(list(self.history.keys()))
        
        for date in sorted_dates:
            daily_data = self.history[date]
            ndt_results = daily_data.get('ndt_results', [])
            
            for row in ndt_results:
                method = str(row.get('검사방법', '')).strip().upper()
                if not method:
                    continue
                    
                if method not in grouped_data:
                    grouped_data[method] = []
                    
                # Inject date into the row dict for convenience
                row_copy = dict(row)
                row_copy['일자'] = date
                grouped_data[method].append(row_copy)
                
        # If no data at all
        if not grouped_data:
            ws = wb.active
            ws.title = "데이터 없음"
            ws['A1'] = "출력할 비파괴검사 결과가 없습니다."
            wb.save(output_path)
            return

        # 1. Create Summary Sheet First
        ws_summary = wb.active
        ws_summary.title = "총괄 집계표"
        self._write_summary_sheet(ws_summary, grouped_data)

        # 2. Create sheets for each method
        for method, rows in grouped_data.items():
            ws = wb.create_sheet(title=method)
            self._write_sheet(ws, method, rows)
            
        wb.save(output_path)
        
    def _write_summary_sheet(self, ws, grouped_data):
        headers = ['업체', '구간(Sec.No)', '라인번호', '관경', '검사방법', '검사 횟수(Joint 카운트)', '물량 누계']
        col_widths = {'A': 15, 'B': 15, 'C': 25, 'D': 12, 'E': 12, 'F': 22, 'G': 25}
        
        for col_idx, width in enumerate(col_widths.values(), start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
            
        # Write Title
        ws.merge_cells('A1:G1')
        title_cell = ws.cell(row=1, column=1, value="용 역 명 : 2026년 중앙지사 열수송관 비파괴검사용역")
        title_cell.font = Font(name='맑은 고딕', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 25

        for col_idx, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header_text)
            cell.font = self.font_bold
            cell.alignment = self.align_center
            cell.fill = self.fill_header
            cell.border = self.border_thin
            
        # Aggregate data
        summary_dict = {}
        for method, rows in grouped_data.items():
            for r in rows:
                company = str(r.get('업체', '')).strip()
                sec = str(r.get('구간', '')).strip()
                line = str(r.get('라인번호', '')).strip()
                size = str(r.get('관경', '')).strip()
                
                key = (company, sec, line, size, method)
                if key not in summary_dict:
                    summary_dict[key] = {'count': 0, 'm': 0.0, 'rt_or': 0, 'rt_re': 0}
                    
                summary_dict[key]['count'] += 1
                
                if method == 'RT':
                    or_val = str(r.get('RT_OR', '')).strip()
                    re_val = str(r.get('RT_RE', '')).strip()
                    try: summary_dict[key]['rt_or'] += int(float(or_val)) if or_val else 0
                    except ValueError: pass
                    try: summary_dict[key]['rt_re'] += int(float(re_val)) if re_val else 0
                    except ValueError: pass
                else:
                    m_val = str(r.get(method, '')).strip().replace(',','')
                    try: summary_dict[key]['m'] += float(m_val) if m_val else 0.0
                    except ValueError: pass
                    
        # Sort and Write
        sorted_keys = sorted(summary_dict.keys())
        row_idx = 3
        for key in sorted_keys:
            company, sec, line, size, method = key
            stats = summary_dict[key]
            
            # 1~5: Group keys
            for i, val in enumerate(key, start=1):
                cell = ws.cell(row=row_idx, column=i, value=val)
                cell.font = self.font_normal
                cell.alignment = self.align_center
                cell.border = self.border_thin
                
            # 6: Count
            cell = ws.cell(row=row_idx, column=6, value=f"{stats['count']} 개")
            cell.font = self.font_normal
            cell.alignment = self.align_center
            cell.border = self.border_thin
            
            # 7: Qty
            cell = ws.cell(row=row_idx, column=7)
            if method == 'RT':
                cell.value = f"OR: {stats['rt_or']} 매 / RE: {stats['rt_re']} 매"
            else:
                cell.value = f"{stats['m']:.4f} m"
                
            cell.font = self.font_normal
            cell.alignment = self.align_center
            cell.border = self.border_thin
            
            row_idx += 1

    def _write_sheet(self, ws, method, rows):
        # Common headers
        headers = ['일자', '순번', '업체', '구간(Sec.No)', '라인번호', 'Joint No.', '관경', '용접사', '구간정보', '결과', '규격']
        
        # Quantity headers
        if method == 'RT':
            qty_headers = ['RT_OR(매)', 'RT_RE(매)']
        else:
            qty_headers = ['물량(m)']
            
        all_headers = headers + qty_headers
        
        # Set column widths
        col_widths = {
            'A': 12, 'B': 6, 'C': 15, 'D': 12, 'E': 25, 'F': 10, 'G': 8, 'H': 15, 'I': 15, 'J': 10, 'K': 10, 'L': 12, 'M': 12
        }
        for col_idx, width in enumerate(col_widths.values(), start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        # Write Title
        num_cols = len(all_headers)
        col_letter = get_column_letter(num_cols)
        ws.merge_cells(f'A1:{col_letter}1')
        title_cell = ws.cell(row=1, column=1, value="용 역 명 : 2026년 중앙지사 열수송관 비파괴검사용역")
        title_cell.font = Font(name='맑은 고딕', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 25

        # Write Headers
        for col_idx, header_text in enumerate(all_headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header_text)
            cell.font = self.font_bold
            cell.alignment = self.align_center
            cell.fill = self.fill_header
            cell.border = self.border_thin

        # Write Data Rows
        for row_idx, row_data in enumerate(rows, start=3):
            col_idx = 1
            
            # Map standard columns
            col_map = {
                '일자': '일자',
                '순번': '순번',
                '업체': '업체',
                '구간(Sec.No)': '구간',
                '라인번호': '라인번호',
                'Joint No.': 'Joint No.',
                '관경': '관경',
                '용접사': '용접사',
                '구간정보': '구간정보',
                '결과': '결과',
                '규격': '규격'
            }
            
            for h in headers:
                key = col_map.get(h, '')
                
                if h == '순번':
                    val = row_idx - 2
                else:
                    val = row_data.get(key, '')
                
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = self.font_normal
                cell.alignment = self.align_center
                cell.border = self.border_thin
                col_idx += 1
                
            # Write quantity columns
            if method == 'RT':
                for k in ['RT_OR', 'RT_RE']:
                    val = row_data.get(k, '')
                    try: val = float(val) if '.' in str(val) else int(val)
                    except ValueError: pass
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = self.font_normal
                    cell.alignment = self.align_center
                    cell.border = self.border_thin
                    col_idx += 1
            else:
                # For PAUT, MT, PT
                val = row_data.get(method, '')
                try: 
                    fval = float(str(val).replace(',',''))
                    cell = ws.cell(row=row_idx, column=col_idx, value=fval)
                    cell.number_format = '0.0000'
                except ValueError:
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    
                cell.font = self.font_normal
                cell.alignment = self.align_center
                cell.border = self.border_thin
                col_idx += 1
