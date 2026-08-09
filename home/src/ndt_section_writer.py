import openpyxl


def safe_write(ws, row, col, value):
    """병합 셀 충돌 없이 안전하게 값 쓰기 (앵커 셀에만 기입)"""
    for merge in ws.merged_cells.ranges:
        if merge.min_row <= row <= merge.max_row and merge.min_col <= col <= merge.max_col:
            ws.cell(row=merge.min_row, column=merge.min_col).value = value
            return
    ws.cell(row=row, column=col).value = value


def find_section_by_title(ws, title_keywords):
    """
    섹션 제목 키워드로 제목 행을 찾고, 그 다음 업체 헤더 행을 탐색합니다.
    Returns: (title_row, header_row, data_start_row)
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value.strip()
                if all(kw in v for kw in title_keywords):
                    title_row = cell.row
                    # 바로 아래에서 '업체' 헤더 찾기
                    for hr in range(title_row + 1, title_row + 10):
                        for col in range(1, 10):
                            hc = ws.cell(row=hr, column=col)
                            if hc.value and '업체' in str(hc.value):
                                return title_row, hr, hr + 1
    return None, None, None


def build_col_map_from_row(ws, header_row):
    """헤더 행을 읽어서 컬럼 매핑을 만듭니다."""
    col_map = {}
    for col in range(1, 30):
        c = ws.cell(row=header_row, column=col)
        if c.value and isinstance(c.value, str):
            v = c.value.strip()
            if '업체' in v:
                col_map['company'] = col
            elif '순번' in v or v in ('번호', 'No'):
                col_map['no'] = col
            elif 'Section' in v or '구간' == v:
                col_map['section'] = col
            elif 'Line No' in v or '라인번호' == v:
                col_map['line_no'] = col
            elif '관경' in v or '구경' in v:
                col_map['pipe_size'] = col
            elif '용접개소' in v or 'Joint' in v:
                col_map['joint'] = col
            elif '용접사' in v:
                col_map['welder'] = col
            elif '규격' == v:
                col_map['shift'] = col
            elif '단위' in v:
                col_map['unit'] = col
            elif '결과' in v:
                col_map['result'] = col
            elif 'PAUT 길이' in v or ('PAUT' in v and '길이' in v):
                col_map['ori'] = col
            elif '촬영매수' in v or '필름' in v:
                col_map['ori'] = col
            elif 'PT 길이' in v:
                col_map['ori'] = col
            elif 'MT 길이' in v:
                col_map['ori'] = col
            elif '비고' in v:
                col_map['note'] = col

    # 서브헤더 행(header_row+1)에서 ORI/RE/TOTAL 찾기
    for col in range(1, 30):
        c = ws.cell(row=header_row + 1, column=col)
        if c.value and isinstance(c.value, str):
            cv = c.value.strip().replace("'", "")
            if cv == 'ORI':
                col_map['ori'] = col
            elif cv == 'RE':
                col_map['re'] = col
            elif 'TOTAL' in cv:
                col_map['total'] = col

    return col_map


def write_records_to_section(ws, records, data_start, col_map):
    """
    records 리스트를 data_start 행부터 기입합니다.
    records: list of dict (업체, 구간, 라인번호, 관경, Joint No., 용접사, ORI, RE, 규격 등)
    """
    if not records:
        return 0

    current_row = data_start
    for i, rec in enumerate(records):
        if 'company' in col_map:
            safe_write(ws, current_row, col_map['company'], rec.get('업체', ''))
        if 'no' in col_map:
            safe_write(ws, current_row, col_map['no'], i + 1)
        if 'section' in col_map:
            safe_write(ws, current_row, col_map['section'], rec.get('구간', ''))
        if 'line_no' in col_map:
            safe_write(ws, current_row, col_map['line_no'], rec.get('라인번호', ''))
        if 'pipe_size' in col_map:
            safe_write(ws, current_row, col_map['pipe_size'], rec.get('관경', ''))
        if 'joint' in col_map:
            safe_write(ws, current_row, col_map['joint'], rec.get('Joint No.', ''))
        if 'welder' in col_map:
            safe_write(ws, current_row, col_map['welder'], rec.get('용접사', ''))
        if 'shift' in col_map:
            safe_write(ws, current_row, col_map['shift'], rec.get('규격', '주간'))
        if 'unit' in col_map:
            safe_write(ws, current_row, col_map['unit'], 'M')
        if 'result' in col_map:
            safe_write(ws, current_row, col_map['result'], rec.get('결과', '합격'))

        ori = float(rec.get('ORI', 0) or 0)
        re_val = float(rec.get('RE', 0) or 0)
        tot = ori + re_val

        if 'ori' in col_map and ori > 0:
            safe_write(ws, current_row, col_map['ori'], round(ori, 4))
        if 're' in col_map and re_val > 0:
            safe_write(ws, current_row, col_map['re'], round(re_val, 4))
        if 'total' in col_map and tot > 0:
            safe_write(ws, current_row, col_map['total'], round(tot, 4))

        current_row += 1

    return len(records)


def extract_records_by_method(history, target_month_str, method_filter):
    """
    history에서 특정 검사방법(PAUT/RT/MT/PT)의 레코드를 추출합니다.
    ORI/RE를 분류하여 list of dict으로 반환합니다.
    """
    raw = []
    for date_key, log_data in history.items():
        if date_key.startswith(target_month_str):
            for r in log_data.get('ndt_results', []):
                method = str(r.get('검사방법', '')).strip().upper()
                if method == method_filter.upper():
                    r['_date'] = date_key
                    raw.append(r)

    raw.sort(key=lambda x: x['_date'])

    groups = {}
    for r in raw:
        key = (
            str(r.get('업체', '')),
            str(r.get('구간', '')),
            str(r.get('라인번호', '')),
            str(r.get('관경', '')),
            str(r.get('Joint No.', '')),
            str(r.get('용접사', ''))
        )
        # 해당 방법의 수치 필드 찾기
        method = method_filter.upper()
        val_fields = {
            'PAUT': ('PAUT',),
            'RT': ('RT_OR', 'RT_ORI', 'RT'),
            'MT': ('MT',),
            'PT': ('PT',),
        }
        val = 0.0
        for field in val_fields.get(method, (method,)):
            raw_val = str(r.get(field, '') or '').strip()
            if raw_val:
                try:
                    v = float(raw_val)
                    if v > 0:
                        val = v
                        break
                except:
                    pass

        if val == 0:
            continue

        shift = str(r.get('규격', '주간')).strip()
        if key not in groups:
            groups[key] = {'ORI': 0.0, 'RE': 0.0, 'shift': shift}
        if groups[key]['ORI'] == 0.0:
            groups[key]['ORI'] += val
        else:
            groups[key]['RE'] += val

    result_list = []
    for (업체, 구간, 라인번호, 관경, joint, 용접사), data in groups.items():
        result_list.append({
            '업체': 업체, '구간': 구간, '라인번호': 라인번호,
            '관경': 관경, 'Joint No.': joint, '용접사': 용접사,
            '규격': data['shift'], '결과': '합격',
            'ORI': data['ORI'], 'RE': data['RE']
        })
    return result_list


def write_all_ndt_sections(ws, history, target_month_str, log_func=print):
    """
    엑셀 시트에서 모든 NDT 섹션(1.2.2~1.2.4, 2.1~2.4)을 찾아 데이터를 기입합니다.
    """
    # 섹션 정의: (title_keywords, method_filter, section_label)
    sections = [
        (['1.2.1', 'PAUT'],      'PAUT', '1.2.1 PAUT'),
        (['1.2.2', 'MT'],        'MT',   '1.2.2 MT'),
        (['1.2.2', '자분'],      'MT',   '1.2.2 자분(MT)'),
        (['1.2.3', 'RT'],        'RT',   '1.2.3 RT'),
        (['1.2.3', '방사선'],    'RT',   '1.2.3 방사선(RT)'),
        (['1.2.4', 'PT'],        'PT',   '1.2.4 PT'),
        (['1.2.4', '침투'],      'PT',   '1.2.4 침투(PT)'),
        (['2.1', 'PAUT'],        'PAUT', '2.1 PAUT'),
        (['2.2', 'MT'],          'MT',   '2.2 MT'),
        (['2.3', 'RT'],          'RT',   '2.3 RT'),
        (['2.4', 'PT'],          'PT',   '2.4 PT'),
    ]

    written_sections = set()
    for kws, method, label in sections:
        if label in written_sections:
            continue
        title_row, header_row, data_start = find_section_by_title(ws, kws)
        if not header_row:
            continue

        col_map = build_col_map_from_row(ws, header_row)
        if not col_map:
            log_func(f"⚠️ {label}: 헤더 컬럼 매핑 실패")
            continue

        records = extract_records_by_method(history, target_month_str, method)
        if not records:
            log_func(f"ℹ️ {label}: {method} 데이터 없음")
            written_sections.add(label)
            continue

        written = write_records_to_section(ws, records, data_start, col_map)
        log_func(f"✅ {label}: {written}건 기입 완료 (row {data_start}~, col_map={col_map})")
        written_sections.add(label)

    return True


# ====================
# 테스트 실행
# ====================
if __name__ == '__main__':
    import json, sys
    sys.path.insert(0, r'c:\Users\jjch2\Desktop\PMI\home\src')
    from paut_writer import find_paut_section, write_paut_data

    history_path = r'c:\Users\jjch2\Desktop\PMI\home\src\daily_work_history.json'
    with open(history_path, 'r', encoding='utf-8') as f:
        history = json.load(f)

    wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\템플릿_최종완성본_V74.xlsx')
    ws = wb.worksheets[0]

    target_month = '2026-08'
    write_all_ndt_sections(ws, history, target_month)

    wb.save(r'C:\Users\jjch2\Desktop\Test_AllNDT.xlsx')
    print("Done! Saved to C:\\Users\\jjch2\\Desktop\\Test_AllNDT.xlsx")
