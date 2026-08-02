import os
import openpyxl
import re

desktop = r'C:\Users\-\OneDrive\바탕 화면'
files = [f for f in os.listdir(desktop) if '산출내역서' in f and f.endswith('.xlsx') and not f.startswith('~$')]

def parse_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    print(f"\n========== {sheet_name} ==========")
    
    current_shift = "일반"
    
    sums = {
        "열배관": {"일반": 0.0, "야간": 0.0, "휴일": 0.0},
        "플랜트(관리소)": {"일반": 0.0, "야간": 0.0, "휴일": 0.0}
    }
    
    for r in range(1, ws.max_row + 1):
        row_str = " ".join([str(ws.cell(row=r, column=c).value) for c in range(1, 20) if ws.cell(row=r, column=c).value is not None])
        
        if '주간' in row_str:
            current_shift = "일반"
        elif '야간' in row_str:
            current_shift = "야간"
        elif '휴일' in row_str:
            current_shift = "휴일"
            
        # Check if row is 열배관 or 관리소
        is_main = '열배관' in row_str
        is_station = '관리소' in row_str
        
        # We need to find the value under "가산~가평" corresponding to the "수량(M)"
        # Let's just print rows that have 열배관 or 관리소 and their columns to see where the data is
        if is_main or is_station:
            cols = []
            for c in range(1, 25):
                val = ws.cell(row=r, column=c).value
                cols.append(str(val) if val is not None else "")
            
            # Print the non-empty columns to debug
            non_empty = [(i+1, v) for i, v in enumerate(cols) if v]
            if len(non_empty) > 3:
                # The total quantity is usually the last numeric value in the row
                print(f"Shift: {current_shift} | Type: {'열배관' if is_main else '관리소'} | Cols: {non_empty}")
                
if files:
    filepath = os.path.join(desktop, files[0])
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for sheet_name in wb.sheetnames:
        if 'UT' in sheet_name or 'PT' in sheet_name:
            parse_sheet(wb, sheet_name)
