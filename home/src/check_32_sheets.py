import openpyxl
import os

file_path = r"C:\Users\jjch2\Desktop\26년07월 월간용역진도보고서.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets = wb.sheetnames
    
    print(f"총 시트 수: {len(sheets)}개\n")
    
    keyword_to_find = "씨지앤대산"
    sheets_with_keyword = []
    
    print("--- 시트 목록 및 프로젝트명 포함 여부 ---")
    for idx, sheet_name in enumerate(sheets, start=1):
        ws = wb[sheet_name]
        found = False
        # Limit the search space slightly for speed, normally project name is at the top
        for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
            for cell_val in row:
                if cell_val and isinstance(cell_val, str) and keyword_to_find in cell_val:
                    found = True
                    break
            if found:
                break
                
        status = "[O] 포함" if found else "[ ] 없음"
        if found:
            sheets_with_keyword.append(sheet_name)
        
        print(f"{idx:02d}. {sheet_name} : {status}")
        
    print("\n[결론] 텍스트 치환이 필요한 시트 목록:")
    print(sheets_with_keyword)

except Exception as e:
    print(f"Error: {e}")
