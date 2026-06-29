import os
import openpyxl

desktop = r'C:\Users\-\OneDrive\바탕 화면'
files = [f for f in os.listdir(desktop) if '산출내역서' in f and f.endswith('.xlsx') and not f.startswith('~$')]

if files:
    filepath = os.path.join(desktop, files[0])
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['6-3. PT']
    
    with open(r'c:\Users\-\OneDrive\바탕 화면\home_new\home\src\pt_dump.txt', 'w', encoding='utf-8') as f:
        f.write("--- 6-3. PT ---\n")
        for r in range(1, min(200, ws.max_row + 1)):
            row_data = []
            for c in range(1, 20):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    row_data.append(str(val).replace('\n', ' '))
                else:
                    row_data.append("")
            if any(row_data):
                f.write(f"R{r}: " + "\t".join(row_data) + "\n")
