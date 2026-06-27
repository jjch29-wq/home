import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

output_path = r"C:\Users\-\OneDrive\바탕 화면\작업승인계획서_NDT전용.xlsx"

# Create a new workbook and select active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "작업승인계획서"

# Define Styles
bold_font = Font(bold=True)
title_font = Font(bold=True, size=16)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

def set_border(ws, min_col, min_row, max_col, max_row):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border

# Column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 15

# --- Document Title ---
ws.merge_cells('A1:E1')
ws['A1'] = "[서울안전건설사무소] (주)OOO 비파괴검사 주요 일일작업"
ws['A1'].font = title_font
ws['A1'].alignment = center_align

# --- Date & Approvals ---
ws.merge_cells('A3:C3')
ws['A3'] = "일자: 2026년 OO월 OO일 O요일"
ws['A3'].font = bold_font

ws['D2'] = "수급업체"
ws['E2'] = "KOGAS"
ws['D3'] = "(인)"
ws['E3'] = "(인)"

for r in range(2, 4):
    for c in range(4, 6):
        ws.cell(row=r, column=c).alignment = center_align
        ws.cell(row=r, column=c).border = thin_border
        ws.cell(row=r, column=c).font = bold_font
ws['D2'].fill = header_fill
ws['E2'].fill = header_fill

# --- Section 1. 투입 현황 ---
ws.merge_cells('A5:E5')
ws['A5'] = "1. 총 투입 현황"
ws['A5'].font = bold_font

# Header row for Section 1
headers_sec1 = ["총 작업 개소", "인원 (계)", "장비 (계)", "RT / 크롤러 투입", "UT / PT / 기타 장비"]
for col, val in enumerate(headers_sec1, start=1):
    cell = ws.cell(row=6, column=col, value=val)
    cell.font = bold_font
    cell.alignment = center_align
    cell.fill = header_fill

values_sec1 = ["00 개소", "00 명", "00 대", "RT조사기 0대, 크롤러 0대", "UT 0대, PT 0세트, 발전기 0대"]
for col, val in enumerate(values_sec1, start=1):
    cell = ws.cell(row=7, column=col, value=val)
    cell.alignment = center_align

set_border(ws, 1, 6, 5, 7)

# --- Section 2. 세부 작업 ---
ws.merge_cells('A9:E9')
ws['A9'] = "2. 팀별 세부 작업 및 안전관리 계획"
ws['A9'].font = bold_font

headers_sec2 = ["구 분", "금 일 작 업 (내용 및 시간)", "주요 위험 요소 (위험성 평가)", "안전관리 중점사항 (대책)", "시공자\n(관리감독자)"]
for col, val in enumerate(headers_sec2, start=1):
    cell = ws.cell(row=10, column=col, value=val)
    cell.font = bold_font
    cell.alignment = center_align
    cell.fill = header_fill

# Row 1 (A Team)
ws['A11'] = "비파괴 A팀 (본관)\n\n(작업개소: 00개소)"
ws['A11'].alignment = center_align

ws['B11'] = "[구간: OO천 ~ OOO천]\n(내용) 30\" 주배관 맞대기 용접부 방사선투과검사(RT)\n※ 작업시간: (08:00~17:00)\n\n[투입 현황]\n인원: 00명\n장비: 조사기 1, 크롤러 1, 차폐막 2"
ws['B11'].alignment = left_align

ws['C11'] = "1. (방사선 피폭) 방사선 투과검사 중 피폭\n2. (추락) 지상 2m 이상 배관 위 검사\n3. (질식) 배관 내부 진입 시 산소 결핍"
ws['C11'].alignment = left_align

ws['D11'] = "1. 콜리메이터 사용, 통제구역 설정/감시자 배치\n2. 고소작업 시 2인 1조 필수, 안전대 체결\n3. 배관내부 인원 진입 금지 (크롤러 대체)"
ws['D11'].alignment = left_align

ws['E11'] = "(서명)"
ws['E11'].alignment = center_align

# Row 2 (B Team)
ws['A12'] = "비파괴 B팀 (관리소)\n\n(작업개소: 00개소)"
ws['A12'].alignment = center_align

ws['B12'] = "[구간: OO관리소 내부]\n(내용) Tie-in 필릿 용접부 초음파(UT) 및 침투(PT)\n※ 작업시간: (08:00~17:00)\n\n[투입 현황]\n인원: 00명\n장비: UT 1, PT 1"
ws['B12'].alignment = left_align

ws['C12'] = "1. (화학물질) PT 용제 취급 시 흡입 위험\n2. (충돌) 좁은 공간 내 타 공정 장비 충돌\n3. (화재) 가연성 가스로 인한 화재"
ws['C12'].alignment = left_align

ws['D12'] = "1. MSDS 비치 및 방독마스크, 장갑 착용\n2. 안전감독관 사전 조율 후 작업 통제\n3. 화기 구역 분리, 소화기 비치"
ws['D12'].alignment = left_align

ws['E12'] = "(서명)"
ws['E12'].alignment = center_align

# Row heights for text wrapping
ws.row_dimensions[11].height = 120
ws.row_dimensions[12].height = 100
ws.row_dimensions[10].height = 30

set_border(ws, 1, 10, 5, 12)

# --- Section 3. 기타 현황 ---
ws.merge_cells('A14:E14')
ws['A14'] = "3. 기타 진행 현황 및 요청사항"
ws['A14'].font = bold_font

headers_sec3 = ["작업 구간", "전일 누계", "금일 계획", "전체 진행률", "기타 작업 현황 및 요청사항"]
for col, val in enumerate(headers_sec3, start=1):
    cell = ws.cell(row=15, column=col, value=val)
    cell.font = bold_font
    cell.alignment = center_align
    cell.fill = header_fill

values_sec3 = ["전체 공구", "000 매/m", "000 매/m", "00.0 %", "- 익일 야간 RT 작업 승인 별도 요청\n- 크롤러 전원 지원 요망"]
for col, val in enumerate(values_sec3, start=1):
    cell = ws.cell(row=16, column=col, value=val)
    if col == 5:
        cell.alignment = left_align
    else:
        cell.alignment = center_align

ws.row_dimensions[16].height = 40
set_border(ws, 1, 15, 5, 16)

# Add Note
ws.merge_cells('A18:E18')
ws['A18'] = "※ 참고: 비파괴검사는 '위험작업'에 속하므로 본 계획서와 함께 [위험성평가표]를 첨부하여 작업 1일 전(D-1)까지 승인을 득해야 합니다."
ws['A18'].font = Font(bold=True, color="FF0000")

# Save workbook
wb.save(output_path)
print(f"Excel file created at: {output_path}")
