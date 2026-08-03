import pandas as pd
import os
import traceback
import tkinter as tk
from tkinter import messagebox, filedialog
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from utils.helpers import normalize_id
import json
import sys
import subprocess
from daily_work_report_manager import DailyWorkReportManager

def export_daily_work_report_impl(self):
    """작업일보를 엑셀 템플릿에 출력합니다."""
    try:
        template_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'resources', 'Template_DailyWorkReport.xlsx')
        if not os.path.exists(template_path):
            template_path = r'c:\Users\jjch2\Desktop\보고서Project PROVIDENCE\Request\PMI\Na-aba\home\resources\Template_DailyWorkReport.xlsx'
        
        if not os.path.exists(template_path):
            messagebox.showerror("오류", "작업일보 템플릿(Template_DailyWorkReport.xlsx)을 찾을 수 없습니다.")
            return

        def _clean_str(v, default=''):
            if v is None: return default
            s = str(v).strip()
            if s.lower() in ('nan', 'none', ''): return default
            return s

        date_val = self.ent_daily_date.get_date()
        site = self.cb_daily_site.get().strip()
        method = self.cb_daily_test_method.get().strip() # Default method from UI
        
        data = {
            'date': date_val,
            'company': _clean_str(self.cb_daily_company.get(), '원자력건설'),
            'project_name': site,
            'standard': _clean_str(self.ent_daily_applied_code.get(), 'KS'),
            'equipment': _clean_str(self.cb_daily_equip.get()),
            'report_no': _clean_str(self.ent_daily_report_no.get()), 
            'inspection_item': _clean_str(self.ent_daily_inspection_item.get()), 
            'inspector': '', 
            'car_no': '', 
            'methods': {},
            'rtk': {},
            'ot_status': [],
            'materials': {}
        }

        if hasattr(self, 'ndt_company_entries') and self.ndt_company_entries:
            company = self.ndt_company_entries[0].get('_company', tk.Variable()).get().strip()
            if company: data['company'] = company

        # DB 데이터 집합 (Option 1: 전체 합산 버전)
        site_records = pd.DataFrame()
        if not self.daily_usage_df.empty:
            try:
                df_copy = self.daily_usage_df.copy()
                
                # [SMART SELECTION] 리스트에서 여러 개를 선택했는지 확인
                selection = self.daily_usage_tree.selection()
                if len(selection) > 1:
                    # 여러 개 선택된 경우, 선택된 항목들만 모아서 출력
                    selected_indices = []
                    for item in selection:
                        tags = self.daily_usage_tree.item(item, 'tags')
                        if tags and tags[0].isdigit():
                            idx = int(tags[0])
                            if idx in df_copy.index:
                                selected_indices.append(idx)
                    if selected_indices:
                        site_records = df_copy.loc[selected_indices]
                        # 현장명/날짜는 첫 번째 선택된 항목 기준으로 업데이트 (데이터 정합성 보장)
                        if 'Site' in site_records.columns: site = str(site_records.iloc[0]['Site']).strip()
                        if 'Date' in site_records.columns: date_val = pd.to_datetime(site_records.iloc[0]['Date']).date()
                    
                if site_records.empty:
                    # 하나만 선택되었거나 선택이 없는 경우: 현재 폼의 날짜/현장 기준으로 전체 합산
                    # [ROBUST MATCHING] 날짜와 현장명 비교 시 공백 등 무시하여 정확도 향상
                    check_date = pd.to_datetime(date_val).date()
                    site_col = 'Site' if 'Site' in df_copy.columns else '현장' if '현장' in df_copy.columns else ''
                    date_col = 'Date' if 'Date' in df_copy.columns else '날짜' if '날짜' in df_copy.columns else ''
                    
                    if site_col and date_col:
                        df_copy['Date_norm'] = pd.to_datetime(df_copy[date_col], errors='coerce').dt.date
                        site_records = df_copy[
                            (df_copy['Date_norm'] == check_date) & 
                            (df_copy[site_col].astype(str).str.strip().str.upper() == str(site).strip().upper())
                        ]
                
                # [NEW] site_records에서 장비명 및 업체명 집계 업데이트
                if not site_records.empty:
                    if '장비명' in site_records.columns:
                        equips = [_clean_str(x) for x in site_records['장비명'].dropna().unique() if _clean_str(x)]
                        if equips: data['equipment'] = ", ".join(equips)
                    
                    comp_col = '업체명' if '업체명' in site_records.columns else ''
                    if comp_col:
                        comps = [_clean_str(x) for x in site_records[comp_col].dropna().unique() if _clean_str(x)]
                        if comps: data['company'] = comps[0] # 첫 번째 유효한 업체명 사용
            except Exception as e:
                print(f"[DEBUG] Site Records Collection Error: {e}")
                pass

        all_vehicles = []
        if hasattr(self, 'vehicle_boxes'):
            for box in self.vehicle_boxes:
                v = box.cb_vehicle_info.get().strip()
                if v and v not in all_vehicles: all_vehicles.append(v)
        if not site_records.empty and '차량번호' in site_records.columns:
            for v in site_records['차량번호'].dropna().unique():
                v_str = str(v).strip()
                if v_str and v_str not in all_vehicles: all_vehicles.append(v_str)
        data['car_no'] = ", ".join(all_vehicles)

        note_texts = []
        # [FIX] A18:S25 메모 영역에는 '비고'가 아닌 '상시 패널(메모)' 내용만 입력되도록 수정
        if hasattr(self, 'main_memo_text'):
            try:
                ui_memo = self.main_memo_text.get('1.0', tk.END).strip()
                if ui_memo:
                    note_texts.append(ui_memo)
            except: pass
                
        data['note'] = "\n".join(note_texts)

        # [NEW] 공사 수행현황 (Section 1) - DB site_records에 있는 모든 방식 집계
        method_col = '검사방법' if '검사방법' in site_records.columns else 'TestMethod' if 'TestMethod' in site_records.columns else ''
        
        if not site_records.empty and method_col:
            for m_name, m_group in site_records.groupby(method_col):
                if not m_name or str(m_name).lower() == 'nan': continue
                m_name_str = str(m_name)
                
                qty_sum = pd.to_numeric(m_group['Usage'], errors='coerce').fillna(0).sum()
                price_max = pd.to_numeric(m_group['단가'], errors='coerce').fillna(0).max()
                travel_sum = pd.to_numeric(m_group['출장비'], errors='coerce').fillna(0).sum()
                total_sum = pd.to_numeric(m_group['검사비'], errors='coerce').fillna(0).sum()
                
                # 단위 정보 찾기 (Unit 또는 단위 컬럼)
                unit_col = 'Unit' if 'Unit' in m_group.columns else '단위' if '단위' in m_group.columns else ''
                unit_val = str(m_group.iloc[0].get(unit_col, '매')) if unit_col else '매'
                
                data['methods'][m_name_str] = {
                    'unit': unit_val,
                    'qty': qty_sum,
                    'price': price_max,
                    'travel': travel_sum,
                    'total': total_sum
                }
        else:
            # Fallback to UI values if DB is empty or column missing
            method = self.cb_daily_test_method.get().strip()
            unit_val = self.cb_daily_unit.get().strip() 
            qty_val = self.ent_daily_test_amount.get().strip()
            price_val = self.ent_daily_unit_price.get().strip()
            travel_val = self.ent_daily_travel_cost.get().strip()
            total_val = self.ent_daily_test_fee.get().strip()
            
            if method:
                data['methods'][method] = {
                    'unit': unit_val,
                    'qty': float(qty_val.replace(',', '')) if qty_val else 0,
                    'price': float(price_val.replace(',', '')) if price_val else 0,
                    'travel': float(travel_val.replace(',', '')) if travel_val else 0,
                    'total': float(total_val.replace(',', '')) if total_val else 0
                }

        # 작업자 및 O/T (간소화 버전)
        # [NEW] 작업자 / OT 정보 DB에서 집계 (현장 탭 기록 기준)
        def _clean_name(n):
            if not n or str(n).lower() in ('nan', 'none', ''): return ''
            text = str(n).strip()
            titles = ['부장', '차장', '과장', '대리', '주임', '기사', '선임', '수석', '책임',
                      '팀장', '이사', '본부장', '실장', '소장', '직장', '반장', '팀원', '계장']
            for t in titles:
                import re as _re
                text = _re.sub(r'[\s/(\[]*' + t + r'[\s)\]]*$', '', text)
                text = _re.sub(r'^[\s/(\[]*' + t + r'[\s)\]]*', '', text)
            return text.strip()

        inspectors = []
        ot_groups = {}  # key: (work_time, ot_amount) -> {names:[], company:''}

        if not site_records.empty:
            company_val = data.get('company', '')
            for _, row in site_records.iterrows():
                for i in range(1, 11):
                    u_key   = 'User'     if i == 1 else f'User{i}'
                    wt_key  = 'WorkTime' if i == 1 else f'WorkTime{i}'
                    ot_key  = 'OT'       if i == 1 else f'OT{i}'

                    name = str(row.get(u_key, '')).strip()
                    if not name or name == 'nan': continue

                    if name not in inspectors:
                        inspectors.append(name)

                    wt  = str(row.get(wt_key, '')).strip()
                    ot_raw = str(row.get(ot_key,  '')).strip()
                    
                    # [FIX] Handle decimals correctly to prevent 26000.0 becoming 260000
                    try:
                        # Try parsing as float first to handle .0
                        oa_num = float(ot_raw.replace(',', '')) if ot_raw and ot_raw != 'nan' else 0
                        oa = str(int(oa_num)) if oa_num > 0 else ''
                    except:
                        # Fallback to digit-only if float parsing fails
                        oa = ''.join(c for c in ot_raw if c.isdigit())

                    if wt == 'nan': wt = ''
                    if not wt and not oa: continue

                    key = (wt, oa)
                    if key not in ot_groups:
                        ot_groups[key] = {'names': [], 'company': company_val}
                    if name not in ot_groups[key]['names']:
                        ot_groups[key]['names'].append(name)

        # Inspector display (titles always stripped)
        disp_insp = [_clean_name(n) for n in inspectors]
        if len(disp_insp) > 3:
            data['inspector'] = ', '.join(disp_insp[:3]) + f" 외 {len(disp_insp) - 3}명"
        else:
            data['inspector'] = ', '.join(disp_insp)

        # Build ot_status list
        data['ot_status'] = []
        if ot_groups:
            for (wt, oa), grp in ot_groups.items():
                names = grp['names']
                name_disp = ', '.join([_clean_name(n) for n in names])
                wt_disp = wt
                
                # Match method for this group if possible
                curr_method = method
                if not site_records.empty and method_col:
                    for _, r_match in site_records.iterrows():
                        if any(str(r_match.get(f'User{k}' if k>1 else 'User', '')).strip() in names for k in range(1, 11)):
                            m_found = str(r_match.get(method_col, '')).strip()
                            if m_found and m_found.lower() != 'nan':
                                curr_method = m_found; break

                data['ot_status'].append({
                    'names':      name_disp,
                    'ot_hours':   wt_disp,
                    'ot_amount':  f'{int(oa):,}' if oa else '',
                    'company':    grp['company'],
                    'method':     curr_method,
                })
        else:
            # Fallback to UI worker groups if DB is empty
            for i in range(1, 11):
                group = getattr(self, f'worker_group{i}', None)
                if group:
                    w_name = group.cb_name.get().strip()
                    w_wt = group.ent_worktime.get().strip()
                    w_ot = group.ent_ot.get().strip()
                    if w_name and (w_wt or w_ot):
                        data['ot_status'].append({
                            'names':      _clean_name(w_name),
                            'ot_hours':   w_wt,
                            'ot_amount':  w_ot,
                            'company':    data.get('company', ''),
                            'method':     method,
                        })

        

        # 2.5 차량 및 안전 점검 수집 (섹션 3)

        data['vehicles'] = []
        seen_vehicles = set()
        if not site_records.empty and '차량번호' in site_records.columns:
            for _, row in site_records.iterrows():
                v_no = str(row.get('차량번호', '')).strip()
                if not v_no or v_no == 'nan': continue
                
                if v_no in seen_vehicles: continue
                seen_vehicles.add(v_no)
                
                v_insp_raw = str(row.get('차량점검', '')).strip()
                v_mileage = str(row.get('주행거리', '')).strip()
                v_remarks = str(row.get('차량비고', '')).strip()
                
                v_parsed = {
                    'vehicle_info': v_no, 
                    'mileage': v_mileage if v_mileage != 'nan' else '', 
                    'remarks': v_remarks if v_remarks != 'nan' else ''
                }
                
                if ':' in v_insp_raw:
                    for pair in v_insp_raw.split('|'):
                        if ':' in pair:
                            k, v = pair.split(':', 1)
                            v_parsed[k] = v
                elif ',' in v_insp_raw or v_insp_raw:
                    for k in v_insp_raw.split(','):
                        k_clean = k.strip()
                        if k_clean:
                            if 'locking' in k_clean: v_parsed[k_clean] = '잠금'
                            elif 'cleaning' in k_clean: v_parsed[k_clean] = '함'
                            else: v_parsed[k_clean] = '양호'
                data['vehicles'].append(v_parsed)
                
        # 폼에 아직 저장되지 않은 내용이 있을 경우 대비
        if not data['vehicles']:
            if hasattr(self, 'vehicle_boxes') and self.vehicle_boxes:
                for box in self.vehicle_boxes:
                    data['vehicles'].append(box.get_data())
            elif hasattr(self, 'vehicle_widget'):
                data['vehicles'].append(self.vehicle_widget.get_data())
                
        # [NEW] N9 셀(car_no)에 들어갈 차량번호만 추출
        car_no_list = []
        for v in data['vehicles']:
            v_no = v.get('vehicle_info', '').strip()
            if v_no:
                car_no_list.append(v_no)
        if car_no_list:
            data['car_no'] = ", ".join(car_no_list)

        # 자재 정보 수집 (NDT 섹션)
        data['selected_material'] = self.cb_daily_material.get().strip()

        # [NEW] RT 품목별 수량 - DB site_records의 MaterialID로 그룹화 + 품목명/규격 조회
        # D열: 품목명(Name), F열: 규격(Spec), M열: 사용수량(Usage)
        if not site_records.empty and 'MaterialID' in site_records.columns:
            # MaterialID 기준으로 Usage 합산 및 Name/Spec 조회
            mat_groups = {}  # key: MaterialID_upper -> {'qty': float, 'name': str, 'spec': str, 'original_id': str}
            for _, row in site_records.iterrows():
                mat_id = str(row.get('MaterialID', '')).strip()
                if not mat_id or mat_id == 'nan':
                    continue
                qty = 0
                try: qty = float(row.get('Usage', 0) or 0)
                except: pass

                mat_id_upper = mat_id.upper()
                if mat_id_upper not in mat_groups:
                    # MaterialID를 사용하여 materials_df에서 실제 품목명과 규격 조회
                    mat_name_val = mat_id
                    mat_spec_val = ''
                    disp_name = ''
                    mat_cat_val = ''
                    
                    if hasattr(self, 'materials_df') and not self.materials_df.empty:
                        try:
                            # 숫자형 ID 대응 (ID가 405.0 등일 수 있음)
                            clean_id = str(mat_id).strip().replace('.0', '')
                            match = self.materials_df[
                                self.materials_df['MaterialID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True) == clean_id
                            ]
                            if not match.empty:
                                # DB에서 값을 가져오되, 품목명 안에 규격이 합쳐져 있을 경우를 대비해 다시 분리
                                raw_name = str(match.iloc[0].get('품목명', mat_name_val)).strip()
                                raw_spec = str(match.iloc[0].get('규격', '')).strip()
                                
                                # [USER REQUEST] Keep 'Carestream' prefix
                                # (Removal logic removed)
                                
                                if '-' in raw_name and (not raw_spec or raw_spec == 'nan' or raw_spec == '' or raw_spec == '자동등록'):
                                    # 품목명에 하이픈이 있고 규격 칸이 비어있거나 무의미한 경우 분리 실행
                                    dash_idx = raw_name.index('-')
                                    mat_name_val = raw_name[:dash_idx].strip()
                                    mat_spec_val = raw_name[dash_idx+1:].strip()
                                else:
                                    mat_name_val = raw_name
                                    mat_spec_val = raw_spec if (raw_spec and str(raw_spec).lower() != 'nan' and raw_spec != '자동등록') else ''
                                
                                # [NEW] Get Category (Classification)
                                mat_cat_val = str(match.iloc[0].get('품목군코드', '')).strip()
                                if not mat_cat_val: mat_cat_val = str(match.iloc[0].iloc[8]) if len(match.columns) > 8 else ''
                        except: pass
                    
                    # [FIX] If we still only have an ID (like '410') and it contains a dash, try splitting it
                    if mat_name_val == mat_id and '-' in mat_id:
                        dash_idx = mat_id.index('-')
                        mat_name_val = mat_id[:dash_idx].strip()
                        mat_spec_val = mat_id[dash_idx+1:].strip()

                    disp_name = f"{mat_name_val}-{mat_spec_val}".upper() if mat_spec_val else mat_name_val.upper()
                    
                    mat_groups[mat_id_upper] = {
                        'qty': 0, 'name': mat_name_val, 'spec': mat_spec_val, 'original_id': mat_id, 'disp': disp_name,
                        'category': mat_cat_val
                    }
                mat_groups[mat_id_upper]['qty'] += qty

            # [REFINEMENT] Secondary Grouping by Display Name to merge identical items
            merged_mats = {} # key: disp_name -> grp_data
            for _, grp in mat_groups.items():
                d_name = grp.get('disp', 'Unknown')
                if d_name not in merged_mats:
                    merged_mats[d_name] = grp.copy()
                else:
                    merged_mats[d_name]['qty'] += grp['qty']

            # Build data['materials'] using merged results
            rt_keys = ['RT T200', 'RT AA400', 'RT Other']
            rt_counter = 0
            
            # [DEBUG] Log all merged materials before filtering
            print(f"DEBUG: All merged materials before RT filtering: {list(merged_mats.keys())}")
            
            # Separate RT items for sequential keying
            # [FIX] Include MT/PT keywords to prevent chemicals from being categorized as RT
            chem_names = [
                'WHITE', 'BLACK', '7C', 'PENETRANT', 'CLEANER', 'DEVELOPER',
                '백색', '흑색', '자분', '침투', '세척', '현상', '페인트',
                'MT', 'PT'
            ]
            rt_merged = {}
            for d_name, grp in merged_mats.items():
                is_chem = any(c in d_name.upper() for c in chem_names)
                # [IMPROVED] Expanded RT detection to catch films by brand and model
                rt_keywords = ['RT ', 'RT-', 'FILM', 'CARESTREAM', 'AGFA', 'FUJI', 'KODAK', 'T200', 'AA400', 'MX125', 'M100']
                cat_upper = str(grp.get('category', '')).upper()
                is_rt_cat = any(k in cat_upper for k in ['FILM', 'RT'])
                is_rt_name = any(k in d_name.upper() for k in rt_keywords)
                
                if (is_rt_cat or is_rt_name) and not is_chem:
                    rt_merged[d_name] = grp

            for d_name, grp in rt_merged.items():
                print(f"DEBUG: Processing RT Item {rt_counter+1}: {d_name}")
                if rt_counter < 3:
                    mat_key = rt_keys[rt_counter]
                else:
                    mat_key = f"RT_ROW_{rt_counter + 1}"
                
                rt_counter += 1
                data['materials'][mat_key] = {
                    'used': float(grp['qty']), 
                    'name': grp['name'], 
                    'spec': grp['spec'], 
                    'is_rt': True, 
                    'category': 'RT'
                }
            
            # [NOTE] Non-RT (Chemicals) are handled separately below

        else:
            # DB 데이터 없으면 UI에서 읽기 (fallback)
            if hasattr(self, 'ndt_company_entries') and self.ndt_company_entries:
                mats = self.ndt_company_entries[0]
                for m_key, mat_name in [('RT T200', 'T200'), ('RT AA400', 'AA400')]:
                    if m_key in mats:
                        val = mats[m_key].get().strip()
                        try: used = int(val) if val else 0
                        except: used = 0
                        data['materials'][m_key] = {'used': used}

        # [NEW] NDT 화학약품 (MT/PT) - DB site_records에서 합산하여 수집
        # [FIX] DT_ 접두어 포함하여 매칭 확장
        chem_db_map = [
            ('MT WHITE',     ['NDT_백색', 'NDT_백색페인트', 'NDT_형광자분', 'NDT_백색페인트_MT', 'DT_백색페인트', 'DT_형광자분']), 
            ('MT 7C-BLACK',  ['NDT_흑색', 'NDT_흑색자분', 'DT_흑색자분']),
            ('PT Penetrant', ['NDT_침투', 'NDT_침투제', 'DT_침투', 'DT_침투제']),
            ('PT Cleaner',   ['NDT_세척', 'NDT_세척제', 'DT_세척', 'DT_세척제']),
            ('PT Developer', ['NDT_현상', 'NDT_현상제', 'DT_현상', 'DT_현상제']),
        ]
        
        db_chem_found = False
        if not site_records.empty:
            # [FIX] DB 컬럼명에 공백이 있을 경우를 대비해 유연하게 매칭 (Normalization)
            actual_cols = list(site_records.columns)
            
            for m_key, db_cols in chem_db_map:
                val_sum = 0
                # [NEW] Check if any rows in site_records have a Material name matching this chemical
                # and if so, sum up their 'Usage' or '수량' columns if they look like chemicals
                for idx, row in site_records.iterrows():
                    m_name = str(row.get('Material', row.get('품목명', ''))).upper()
                    # Check if this row's material matches any of our chemical patterns
                    is_match = False
                    for col_pattern in db_cols:
                        pattern_norm = col_pattern.replace('NDT_', '').replace('DT_', '').replace(' ', '').upper()
                        if pattern_norm in m_name.replace(' ', ''):
                            is_match = True
                            break
                    
                    if is_match:
                        # If it matches, try 'Usage' first, then '수량' if it looks small (like cans)
                        u_val = pd.to_numeric(row.get('Usage'), errors='coerce')
                        if pd.isna(u_val): u_val = 0
                        val_sum += u_val
                        db_chem_found = True

                # Also check specific columns if they exist as fallback/alternative
                for col_pattern in db_cols:
                    pattern_norm = col_pattern.replace(' ', '').upper()
                    found_col = None
                    for actual_col in actual_cols:
                        if str(actual_col).replace(' ', '').upper() == pattern_norm:
                            found_col = actual_col
                            break
                    
                    if found_col:
                        try:
                            val_sum += int(pd.to_numeric(site_records[found_col], errors='coerce').fillna(0).sum())
                            db_chem_found = True
                        except: pass
                
                if val_sum > 0:
                    # [FIX] Use float for chemical quantities
                    data['materials'][m_key] = {'used': float(val_sum)}
                    db_chem_found = True
        
        # Debug gathered materials data
        print(f"DEBUG: Gathered Materials Data for Report: {data['materials']}")
        
        # DB에 데이터가 없으면 UI 위젯에서 읽기 (Fallback)
        if not db_chem_found and hasattr(self, 'ndt_company_entries') and self.ndt_company_entries:
            entries_dict = self.ndt_company_entries[0]
            mat_keys = [k for k in entries_dict.keys() if k != '_company']
            fallback_order = [
                ('MT WHITE', 0), ('MT 7C-BLACK', 1), 
                ('PT Penetrant', 3), ('PT Cleaner', 4), ('PT Developer', 5)
            ]
            for m_key, idx in fallback_order:
                if idx < len(mat_keys):
                    try:
                        val = entries_dict[mat_keys[idx]].get().strip()
                        used = int(val) if val else 0
                        if used > 0: data['materials'][m_key] = {'used': used}
                    except: pass

        # 2.5 OT Status Gathering
        data['ot_status'] = []
        
        def clean_val(v):
            if pd.isna(v) or str(v).lower() == 'nan': return ""
            val = str(v).strip()
            # Strip common titles from both END and START
            titles = ["부장", "차장", "과장", "대리", "주임", "계장", "사원", "반장", "기사"]
            for title in titles:
                if val.endswith(title): val = val[:-len(title)].strip()
                if val.startswith(title): val = val[len(title):].strip()
            return val

        # 1) Priority: Check DB (Site Records) first for "Saved" data
        if not site_records.empty:
            print(f"DEBUG: --- ALL DB COLUMNS: {site_records.columns.tolist()} ---")
            for idx, row in site_records.iterrows():
                # [FINAL FIX] Scan all 10 user columns to ensure no one is missed (e.g. User2, User3...)
                for i in range(1, 11):
                    u_key = 'User' if i == 1 else f'User{i}'
                    wt_key = 'WorkTime' if i == 1 else f'WorkTime{i}'
                    oa_key = 'OT' if i == 1 else f'OT{i}'
                    
                    worker_name = clean_val(row.get(u_key, row.get('작업자' if i==1 else f'작업자{i}', '')))
                    if not worker_name or worker_name == "": continue
                    
                    ot_val = clean_val(row.get(wt_key, ''))
                    if not ot_val:
                        ot_val = clean_val(row.get('작업시간' if i==1 else f'작업시간{i}', ''))
                    
                    # Fallback for WorkTime: If specific UserX's WorkTime is missing, try general WorkTime column
                    if not ot_val:
                        ot_val = clean_val(row.get('WorkTime', row.get('작업시간', '')))
                    
                    if not ot_val:
                        # Scan all columns for time-like strings if explicit names fail
                        for col_name in row.index:
                            potential_val = str(row[col_name])
                            if any(k in potential_val for k in ["09:00", "24:00", "(주야간)", "~"]):
                                ot_val = potential_val.strip()
                                break
                    
                    if not ot_val:
                        ot_val = clean_val(row.get('OT시간' if i==1 else f'OT시간{i}', ''))
                        
                    # For amount, 'OT' seems to have it in this DB
                    ot_amount = clean_val(row.get(oa_key, row.get('OT금액' if i==1 else f'OT금액{i}', '')))
                    if not ot_amount and i > 1:
                        # Fallback to general OT column if specific one is empty
                        ot_amount = clean_val(row.get('OT', row.get('OT금액', '')))
                    
                    data['ot_status'].append({
                        'names': worker_name,
                        'company': clean_val(row.get('업체명', '')),
                        'method': clean_val(row.get('검사방법', '')),
                        'ot_hours': ot_val,
                        'ot_amount': ot_amount
                    })
            print(f"DEBUG: Gathered {len(data['ot_status'])} OT items from DB")

        # 2) Fallback: If DB is empty, check UI entries
        if not data['ot_status']:
            for i in range(1, 11):
                group = getattr(self, f'worker_group{i}', None)
                if group:
                    try:
                        name = group.get_worker().strip()
                        if name and name.lower() != 'nan':
                            ot_val = ""
                            if hasattr(group, 'ent_ot'): ot_val = group.ent_ot.get().strip()
                            elif hasattr(group, 'cb_ot'): ot_val = group.cb_ot.get().strip()
                                
                            data['ot_status'].append({
                                'names': name,
                                'company': self.cb_daily_company.get().strip(),
                                'method': self.cb_daily_test_method.get().strip(),
                                'ot_hours': ot_val,
                                'ot_amount': '' 
                            })
                    except: pass
            if data['ot_status']:
                print(f"DEBUG: Gathered {len(data['ot_status'])} OT items from UI (Fallback)")
        
        print(f"DEBUG: FINAL OT Status for report: {len(data['ot_status'])} items")

        data['rtk'] = {}
        rtk_cats = {
            '센터미스': 'center_miss', '농도': 'density', '마킹미스': 'marking_miss',
            '필름마크': 'film_mark', '취급부주의': 'handling', '고객불만': 'customer_complaint', '기타': 'etc'
        }
        rtk_total = 0
        if not site_records.empty:
            # Get clean column names for matching
            import re as _re
            def clean_name(s): return _re.sub(r'[^A-Z가-힣0-9]', '', str(s).upper())
            
            actual_cols = site_records.columns.tolist()
            clean_cols = [clean_name(c) for c in actual_cols]
            
            for kor_key, eng_key in rtk_cats.items():
                db_val = 0
                target_clean = clean_name(kor_key)
                rtk_target_clean = clean_name(f"RTK_{kor_key}")
                
                found_col = None
                for i, c_clean in enumerate(clean_cols):
                    if c_clean == target_clean or c_clean == rtk_target_clean:
                        found_col = actual_cols[i]
                        break
                
                if found_col:
                    try:
                        # Sum up and handle potential string/NaN values
                        series = pd.to_numeric(site_records[found_col], errors='coerce').fillna(0)
                        db_val = int(series.sum())
                        print(f"DEBUG: Found RTK Col '{found_col}' for '{kor_key}', sum={db_val}")
                    except Exception as e:
                        print(f"DEBUG: Error summing RTK {kor_key}: {e}")
                
                data['rtk'][kor_key] = db_val
                rtk_total += db_val
        
        # [CRITICAL FIX] If DB result is 0 or records empty, fallback to UI entries
        if rtk_total == 0:
            for kor_key, eng_key in rtk_cats.items():
                val = 0
                # Try direct widget attribute first
                widget = getattr(self, f"ent_rtk_{eng_key}", None)
                if widget:
                    try: val = int(widget.get().strip() or 0)
                    except: pass
                # Then try the rtk_entries dictionary as backup
                elif hasattr(self, 'rtk_entries') and kor_key in self.rtk_entries:
                    try: val = int(self.rtk_entries[kor_key].get().strip() or 0)
                    except: pass
                
                data['rtk'][kor_key] = val
                rtk_total += val
                
        data['rtk_total'] = rtk_total

        default_filename = f"한국지역난방 중앙지사_{date_val.strftime('%Y%m%d')}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="작업일보 저장",
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )

        if not save_path:
            return

        # 4. 리포트 생성
        mapping = self.load_report_mapping()
        manager = DailyWorkReportManager(template_path)
        manager.generate_report(data, save_path, custom_mapping=mapping)

        messagebox.showinfo("성공", f"작업일보가 생성되었습니다:\n{save_path}")
        
        # 생성된 파일 열기
        if messagebox.askyesno("확인", "생성된 작업일보를 지금 여시겠습니까?"):
            os.startfile(save_path)

    except PermissionError:
        messagebox.showerror("오류", f"파일 접근 오류: '{os.path.basename(save_path)}' 파일이 다른 프로그램(엑셀 등)에서 열려 있어 저장할 수 없습니다.\n파일을 닫고 다시 시도해 주세요.")
    except Exception as e:
        traceback.print_exc()
        messagebox.showerror("오류", f"작업일보 생성 중 오류 발생: {e}")


def export_materials_impl(self):
    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        initialfile="Materials_Export.xlsx",
        title="자재 목록 저장",
        filetypes=[("Excel files", "*.xlsx")]
    )
    
    if save_path:
        try:
            self.materials_df.to_excel(save_path, index=False)
            messagebox.showinfo("완료", "자재 목록이 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패: {e}")


def save_df_to_excel_autofit_impl(self, df, save_path, sheet_name='Sheet1'):
    """Save a DataFrame to Excel with automatic column width adjustment (AutoFit)"""
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        
        for idx, col in enumerate(df.columns):
            # Calculate max length of values in column + header
            # We handle Korean characters by assuming they take ~2 units of width
            def get_display_width(s):
                width = 0
                for char in str(s):
                    if ord(char) > 127: # Non-ASCII (Korean, etc.)
                        width += 2
                    else:
                        width += 1
                return width

            series = df[col].astype(str)
            # Filter out empty strings/NAs for max calc
            lengths = series.apply(get_display_width)
            max_val_len = lengths.max() if not lengths.empty else 0
            header_len = get_display_width(col)
            
            # Final width with padding - Cap at a reasonable width to force wrapping
            final_width = min(max(max_val_len, header_len) + 2, 15)
            
            # Map index to column letter
            # column_letter property is available in openpyxl cells
            col_letter = worksheet.cell(row=1, column=idx+1).column_letter
            worksheet.column_dimensions[col_letter].width = final_width

        # --- Page Setup for Printing ---
        # Set to Portrait orientation (User request)
        worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
        
        # Enable Fit to Page width (Scale everything to fit horizontally)
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0 # Automatic heights (multiple pages if long)
        
        # Enable Wrap Text for all cells to use vertical space instead of horizontal width
        from openpyxl.styles import Alignment
        wrap_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = wrap_alignment

        # Optional: Set small margins to maximize space (units are inches)
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.5
        worksheet.page_margins.bottom = 0.5





def export_central_daily_work_report_impl(self):
    """작업일보를 중앙지사 전용 v20 양식에 맞추어 엑셀로 출력합니다."""
    try:
        import pandas as pd
        import openpyxl
        from datetime import datetime
        import tkinter as tk
        from tkinter import messagebox, filedialog

        template_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'resources', 'Template_Central_DailyWorkReport.xlsx')
        if not os.path.exists(template_path):
            messagebox.showerror("오류", "중앙지사 전용 작업일보 템플릿(Template_Central_DailyWorkReport.xlsx)을 찾을 수 없습니다.")
            return

        def _clean_str(v, default=''):
            if v is None: return default
            s = str(v).strip()
            if s.lower() in ('nan', 'none', ''): return default
            return s
            
        def _clean_name(n):
            if not n or str(n).lower() in ('nan', 'none', ''): return ''
            text = str(n).strip()
            titles = ['부장', '차장', '과장', '대리', '주임', '기사', '선임', '수석', '책임',
                      '팀장', '이사', '본부장', '실장', '소장', '직장', '반장', '팀원', '계장']
            for t in titles:
                import re as _re
                text = _re.sub(r'[\s/(\[]*' + t + r'[\s)\]]*$', '', text)
                text = _re.sub(r'^[\s/(\[]*' + t + r'[\s)\]]*', '', text)
            return text.strip()

        date_val = self.ent_daily_date.get_date()
        site = self.cb_daily_site.get().strip()

        # DB 데이터 집합 
        df_copy = pd.DataFrame()
        if not self.daily_usage_df.empty:
            df_copy = self.daily_usage_df.copy()
            site_col = 'Site' if 'Site' in df_copy.columns else '현장' if '현장' in df_copy.columns else ''
            date_col = 'Date' if 'Date' in df_copy.columns else '날짜' if '날짜' in df_copy.columns else ''
            
            if site_col and date_col:
                df_copy['Date_norm'] = pd.to_datetime(df_copy[date_col], errors='coerce').dt.date
                check_date = pd.to_datetime(date_val).date()
                
                # 금일 기록
                today_records = df_copy[
                    (df_copy['Date_norm'] == check_date) & 
                    (df_copy[site_col].astype(str).str.strip().str.upper() == str(site).strip().upper())
                ]
                
                # 전일 누계 기록 (과거 기록 모두 합산)
                past_records = df_copy[
                    (df_copy['Date_norm'] < check_date) & 
                    (df_copy[site_col].astype(str).str.strip().str.upper() == str(site).strip().upper())
                ]
            else:
                today_records = pd.DataFrame()
                past_records = pd.DataFrame()
        else:
            today_records = pd.DataFrame()
            past_records = pd.DataFrame()

        # 물량 집계 로직
        def calc_method_qty(records, method_key, base_method):
            if records.empty: return 0
            method_col = '검사방법' if '검사방법' in records.columns else 'TestMethod' if 'TestMethod' in records.columns else ''
            if not method_col: return 0
            
            qty = 0
            for _, row in records.iterrows():
                m = str(row.get(method_col, '')).upper().strip()
                if base_method in m:
                    # 야간 확인
                    is_night = '야간' in m or 'NIGHT' in m
                    target_is_night = '_N' in method_key
                    
                    # 관경 확인 (PAUT, RT의 경우)
                    size_match = True
                    if '300A' in method_key and not ('300' in m or '400' in m or '500' in m): size_match = False
                    if '250A' in method_key and not ('250' in m): size_match = False
                    if '200A' in method_key and not ('200' in m): size_match = False
                    if '150A' in method_key and not ('150' in m or '100' in m): size_match = False
                    if '80A' in method_key and not ('80' in m or '65' in m or '50' in m or '40' in m): size_match = False
                    
                    if size_match and (is_night == target_is_night):
                        qty += pd.to_numeric(row.get('Usage', 0), errors='coerce')
            return float(qty) if pd.notna(qty) else 0

        # 장비 수집
        equips = []
        if not today_records.empty and '장비명' in today_records.columns:
            equips = [_clean_str(x) for x in today_records['장비명'].dropna().unique() if _clean_str(x)]
        if not equips:
            equips = [_clean_str(self.cb_daily_equip.get())]
            
        equip_str = ", ".join(filter(None, equips))

        # 인원 수집
        inspectors = []
        managers = []
        if not today_records.empty:
            for _, row in today_records.iterrows():
                for i in range(1, 11):
                    u_key = 'User' if i == 1 else f'User{i}'
                    name = str(row.get(u_key, '')).strip()
                    if name and name != 'nan':
                        clean_n = _clean_name(name)
                        if '대리인' in name or '안전' in name or '관리' in name:
                            if clean_n not in managers: managers.append(clean_n)
                        else:
                            if clean_n not in inspectors: inspectors.append(clean_n)
                            
        if not inspectors and not managers:
            managers_raw = ""
            managers = [_clean_name(x.strip()) for x in managers_raw.split(',') if x.strip()]
            for i in range(1, 11):
                group = getattr(self, f'worker_group{i}', None)
                if group:
                    w_name = _clean_name(group.get_worker().strip())
                    if w_name and w_name not in inspectors: inspectors.append(w_name)

        # 엑셀 오픈
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # 상단 데이터 맵핑
        ws['A4'] = f"검사일자 : {date_val.strftime('%Y년 %m월 %d일')}           날 씨 : "
        
        # 1. 작업 물량 및 누계 현황 맵핑 (PAUT: 행 7~11, RT: 13~16, MT: 18~19, PT: 20~21)
        # [행번호, method_key, base_method]
        mapping_rows = [
            [7, 'PAUT_300A_D', 'PAUT'],
            [8, 'PAUT_300A_N', 'PAUT'],
            [9, 'PAUT_250A', 'PAUT'],
            [10, 'PAUT_200A_D', 'PAUT'],
            [11, 'PAUT_200A_N', 'PAUT'],
            # 12 is PAUT 소계
            [13, 'RT_150A_D', 'RT'],
            [14, 'RT_150A_N', 'RT'],
            [15, 'RT_80A_D', 'RT'],
            [16, 'RT_80A_N', 'RT'],
            # 17 is RT 소계
            [18, 'MT_D', 'MT'],
            [19, 'MT_N', 'MT'],
            [20, 'PT_D', 'PT'],
            [21, 'PT_N', 'PT']
        ]

        for r, m_key, base_m in mapping_rows:
            today_qty = calc_method_qty(today_records, m_key, base_m)
            past_qty = calc_method_qty(past_records, m_key, base_m)
            total_qty = today_qty + past_qty
            
            # 예상량(C열) 가져오기
            exp_val = ws.cell(row=r, column=3).value
            try: exp_val = float(exp_val) if exp_val else 0
            except: exp_val = 0
            
            # 전일 누계 (D열)
            ws.cell(row=r, column=4).value = past_qty if past_qty > 0 else ''
            # 금일 작업 (E열)
            ws.cell(row=r, column=5).value = today_qty if today_qty > 0 else ''
            # 총 누계 (F열)
            ws.cell(row=r, column=6).value = total_qty if total_qty > 0 else ''
            # 공정률 (G열)
            if exp_val > 0 and total_qty > 0:
                ws.cell(row=r, column=7).value = round((total_qty / exp_val) * 100, 1)
            else:
                ws.cell(row=r, column=7).value = ''

        # 소계 엑셀 수식은 엑셀 자체에 없으므로 파이썬에서 계산하여 넣음
        paut_past = sum([calc_method_qty(past_records, m, 'PAUT') for m in ['PAUT_300A_D', 'PAUT_300A_N', 'PAUT_250A', 'PAUT_200A_D', 'PAUT_200A_N']])
        paut_today = sum([calc_method_qty(today_records, m, 'PAUT') for m in ['PAUT_300A_D', 'PAUT_300A_N', 'PAUT_250A', 'PAUT_200A_D', 'PAUT_200A_N']])
        ws.cell(row=12, column=4).value = paut_past if paut_past > 0 else ''
        ws.cell(row=12, column=5).value = paut_today if paut_today > 0 else ''
        ws.cell(row=12, column=6).value = (paut_past + paut_today) if (paut_past + paut_today) > 0 else ''
        
        rt_past = sum([calc_method_qty(past_records, m, 'RT') for m in ['RT_150A_D', 'RT_150A_N', 'RT_80A_D', 'RT_80A_N']])
        rt_today = sum([calc_method_qty(today_records, m, 'RT') for m in ['RT_150A_D', 'RT_150A_N', 'RT_80A_D', 'RT_80A_N']])
        ws.cell(row=17, column=4).value = rt_past if rt_past > 0 else ''
        ws.cell(row=17, column=5).value = rt_today if rt_today > 0 else ''
        ws.cell(row=17, column=6).value = (rt_past + rt_today) if (rt_past + rt_today) > 0 else ''
        
        # 3. 장비 및 인원 맵핑 (v20 양식 구조 맞춤)
        ws.cell(row=7, column=12).value = equip_str # L7 (L7:M7 병합)

        ws['Q7'] = len(managers) if managers else ''
        ws['Q8'] = ", ".join(managers) if managers else ''
        
        ws['R7'] = len(inspectors) if inspectors else ''
        ws['R8'] = ", ".join(inspectors) if inspectors else ''

        # 결과 저장
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f"{site}_검사일보_{date_val.strftime('%y%m%d')}.xlsx"
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            initialfile=default_name, 
            filetypes=[("Excel File", "*.xlsx")], 
            title="중앙지사 검사일보 저장"
        )
        
        if filepath:
            wb.save(filepath)
            wb.close()
            messagebox.showinfo("완료", f"중앙지사 엑셀 작업일보 생성이 완료되었습니다.\n(2번 비파괴검사결과서는 엑셀에서 직접 기입해 주세요)\n\n{filepath}")
            os.startfile(filepath)
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        messagebox.showerror("오류", f"엑셀 생성 중 오류가 발생했습니다: {str(e)}")
