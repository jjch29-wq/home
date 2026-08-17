import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.cell.cell import Cell
import os
import json
import datetime
from collections import defaultdict
import re

class MonthlyReportManager:
    WELDER_NAMES = {
        'W-2023-A-10': '이신희',
        'W-2023-A-13': '선성문',
        'W-2023-A-25': '이종근',
    }

    def __init__(self, template_path):
        self.template_path = template_path
        # 템플릿의 각 표 시작 위치 (헤더 기준)
        # 이 위치는 템플릿_최종완성본_V70 기준이며, 동적 삽입 시 밀릴 수 있으므로 
        # 위에서부터 아래로 순차적으로 삽입(insert_rows)하면서 작업해야 합니다.
        self.table_markers = {
            'qty': 382, # 1.1 비파괴검사 물량표
            'paut_1': 403, # 1.2.1 PAUT
            'mt_1': 412, # 1.2.2 MT
            'rt_1': 429, # 1.2.3 RT
            'pt_1': 436, # 1.2.4 PT
            'paut_2': 448, # 2.1 PAUT 세부
            'mt_2': 481, # 2.2 MT 세부
            'rt_2': 490, # 2.3 RT 세부
            'pt_2': 497 # 2.4 PT 세부
        }
        
        self.font_normal = Font(name='맑은 고딕', size=10)
        self.align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin')
        self.border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _insert_rows_safely(self, ws, insert_idx, amount):
        """
        openpyxl의 병합 셀 깨집 버그를 우회하여 안전하게 행을 삽입합니다.
        삽입 위치 아래의 병합 셀 구조를 보존합니다.
        """
        from openpyxl.worksheet.cell_range import CellRange
        from openpyxl.worksheet.pagebreak import Break
        # 1. 모든 병합 정보 백업
        old_merges = [str(m) for m in ws.merged_cells.ranges]
        # 2. 병합 정보 초기화 (에러 원인 제거)
        ws.merged_cells.ranges = []
        # 3. 안전하게 빈 행 삽입
        ws.insert_rows(insert_idx, amount)
        # 4. 백업된 병합 정보를 위치에 맞게 수정하여 다시 적용
        for m_str in old_merges:
            cr = CellRange(m_str)
            if cr.min_row >= insert_idx:
                cr.shift(row_shift=amount)   # 삽입 위치 아래 병합셀 밀어냄
            elif cr.min_row < insert_idx <= cr.max_row:
                cr.expand(down=amount)       # 삽입 위치에 걸치는 병합셀 늘림
            ws.merge_cells(str(cr))

        # 5. 페이지 브레이크(로고/쪽번호 위치) 보정
        if ws.row_breaks and ws.row_breaks.brk:
            new_breaks = []
            for brk in ws.row_breaks.brk:
                if brk.id >= insert_idx:
                    new_breaks.append(Break(id=brk.id + amount))
                else:
                    new_breaks.append(brk)
            ws.row_breaks.brk = new_breaks

        # 6. 인쇄 영역(print_area) 보정
        if ws.print_area:
            try:
                pa = CellRange(str(ws.print_area).replace('$', ''))
                if pa.max_row >= insert_idx:
                    pa.max_row += amount
                    ws.print_area = str(pa)
            except Exception:
                pass
    def _delete_rows_safely(self, ws, delete_idx, amount):
        """
        openpyxl의 병합 셀 깨집 버그를 우회하여 안전하게 행을 삭제합니다.
        """
        from openpyxl.worksheet.cell_range import CellRange
        from openpyxl.worksheet.pagebreak import Break
        # 1. 삭제 범위에 완전히 포함되지 않는 병합 보존
        old_merges = []
        for m in ws.merged_cells.ranges:
            if m.min_row >= delete_idx and m.max_row < delete_idx + amount:
                continue  # 삭제될 행 안에만 있는 병합은 제거
            old_merges.append(str(m))

        ws.merged_cells.ranges = []
        # 2. 행 삭제
        ws.delete_rows(delete_idx, amount)
        # 3. 병합 정보 복원 (삭제 위치 기준으로 조정)
        for m_str in old_merges:
            cr = CellRange(m_str)
            if cr.min_row >= delete_idx + amount:
                cr.shift(row_shift=-amount)   # 삭제 아래 병합셀 당김
            elif cr.max_row >= delete_idx + amount:
                cr.max_row -= amount          # 범위에 걸치는 병합셀 축소
            try:
                ws.merge_cells(str(cr))
            except Exception:
                pass

        # Adjust page breaks after deleted rows.
        if ws.row_breaks and ws.row_breaks.brk:
            new_breaks = []
            for brk in ws.row_breaks.brk:
                if brk.id >= delete_idx + amount:
                    new_breaks.append(Break(id=brk.id - amount))
                elif brk.id < delete_idx:
                    new_breaks.append(brk)
            ws.row_breaks.brk = new_breaks

    def _repair_merged_right_borders(self, ws):
        """Restore right edges that openpyxl can drop from merged template cells.

        Excel stores a merged range's visible outline across both the anchor cell
        and the cells on the perimeter.  After a load/save cycle, the right edge
        can disappear when the template only carries the outline on the anchor.
        Reapply that edge to the full right perimeter before saving.
        """
        import copy

        for merged in list(ws.merged_cells.ranges):
            anchor = ws.cell(row=merged.min_row, column=merged.min_col)

            # Prefer an existing right edge.  For framed merged cells whose right
            # edge was already lost, the matching left edge is the intended style.
            right_side = anchor.border.right
            if not right_side.style:
                for row in range(merged.min_row, merged.max_row + 1):
                    candidate = ws.cell(row=row, column=merged.max_col).border.right
                    if candidate.style:
                        right_side = candidate
                        break

            is_framed = bool(
                anchor.border.left.style
                and (anchor.border.top.style or anchor.border.bottom.style)
            )
            if not right_side.style and is_framed:
                right_side = anchor.border.left

            if not right_side.style:
                continue

            # Keep the anchor edge as well as every cell on the visible perimeter.
            targets = [anchor]
            targets.extend(
                ws.cell(row=row, column=merged.max_col)
                for row in range(merged.min_row, merged.max_row + 1)
            )
            for cell in targets:
                border = copy.copy(cell.border)
                border.right = copy.copy(right_side)
                cell.border = border

    def _insert_continuation_header(self, ws, insert_row, source_row, row_count=6):
        """Insert a copy of an existing page header before continued table rows."""
        import copy

        max_col = ws.max_column
        cell_snapshot = []
        for r_offset in range(row_count):
            row_data = []
            for col in range(1, max_col + 1):
                cell = ws.cell(row=source_row + r_offset, column=col)
                row_data.append({
                    'value': cell.value,
                    'style': copy.copy(cell._style),
                    'number_format': cell.number_format,
                })
            cell_snapshot.append(row_data)

        row_heights = [
            ws.row_dimensions[source_row + offset].height
            for offset in range(row_count)
        ]
        source_merges = []
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row >= source_row and merged.max_row < source_row + row_count:
                source_merges.append((
                    merged.min_row - source_row, merged.min_col,
                    merged.max_row - source_row, merged.max_col,
                ))

        # Images are drawing objects, so insert_rows() neither moves nor copies
        # them.  Keep snapshots of images anchored in the source header.
        source_images = []
        for image in list(ws._images):
            anchor = getattr(image, 'anchor', None)
            if not hasattr(anchor, '_from'):
                continue
            image_row = anchor._from.row + 1
            if source_row <= image_row < source_row + row_count:
                source_images.append((image, copy.deepcopy(anchor)))

        self._insert_rows_safely(ws, insert_row, row_count)

        # Move existing drawings below the insertion together with their cells.
        for image in list(ws._images):
            anchor = getattr(image, 'anchor', None)
            if not hasattr(anchor, '_from') or anchor._from.row + 1 < insert_row:
                continue
            anchor._from.row += row_count
            if hasattr(anchor, 'to'):
                anchor.to.row += row_count

        for r_offset, row_data in enumerate(cell_snapshot):
            ws.row_dimensions[insert_row + r_offset].height = row_heights[r_offset]
            for col, saved in enumerate(row_data, start=1):
                cell = ws.cell(row=insert_row + r_offset, column=col)
                cell.value = saved['value']
                cell._style = copy.copy(saved['style'])
                cell.number_format = saved['number_format']

        for min_ro, min_col, max_ro, max_col_idx in source_merges:
            ws.merge_cells(
                start_row=insert_row + min_ro, start_column=min_col,
                end_row=insert_row + max_ro, end_column=max_col_idx,
            )

        # Duplicate source-header logos/images at the continuation header.
        image_row_shift = insert_row - source_row
        for source_image, saved_anchor in source_images:
            from io import BytesIO
            from openpyxl.drawing.image import Image as XLImage

            image_buffer = BytesIO()
            image_ref = source_image.ref
            if hasattr(image_ref, 'copy') and hasattr(image_ref, 'save'):
                pil_copy = image_ref.copy()
                image_format = getattr(image_ref, 'format', None) or 'PNG'
                pil_copy.save(image_buffer, format=image_format)
            elif isinstance(image_ref, (str, os.PathLike)):
                with open(image_ref, 'rb') as image_file:
                    image_buffer.write(image_file.read())
            else:
                old_position = image_ref.tell() if hasattr(image_ref, 'tell') else None
                if hasattr(image_ref, 'seek'):
                    image_ref.seek(0)
                image_buffer.write(image_ref.read())
                if old_position is not None and hasattr(image_ref, 'seek'):
                    image_ref.seek(old_position)

            image_buffer.seek(0)
            new_image = XLImage(image_buffer)
            new_image.width = source_image.width
            new_image.height = source_image.height
            # Retain the stream for the whole workbook-save lifecycle.
            new_image._continuation_stream = image_buffer
            new_image.anchor = copy.deepcopy(saved_anchor)
            new_image.anchor._from.row += image_row_shift
            if hasattr(new_image.anchor, 'to'):
                new_image.anchor.to.row += image_row_shift
            ws.add_image(new_image)

    def _insert_table_continuation(self, ws, insert_row, document_header_row,
                                   table_header_row):
        """Start a continuation page with the document and table headers.

        ``table_header_row`` is the first of the table's two column-header
        rows; the subsection title is the row immediately above it.  The
        resulting continuation block is eleven rows high:

        * seven rows of the repeated document header;
        * one blank spacer row below the document header;
        * one subsection-title row;
        * two table-column-header rows.

        Existing content at and below ``insert_row`` is moved down intact.
        """
        import copy

        title_row = table_header_row - 1
        source_rows = range(title_row, table_header_row + 2)
        max_col = ws.max_column

        cell_snapshot = []
        for source_row in source_rows:
            row_snapshot = []
            for col in range(1, max_col + 1):
                cell = ws.cell(row=source_row, column=col)
                row_snapshot.append({
                    'value': cell.value,
                    'style': copy.copy(cell._style),
                    'number_format': cell.number_format,
                })
            cell_snapshot.append(row_snapshot)

        row_heights = [ws.row_dimensions[row].height for row in source_rows]
        source_merges = []
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row >= title_row and merged.max_row <= table_header_row + 1:
                source_merges.append((
                    merged.min_row - title_row, merged.min_col,
                    merged.max_row - title_row, merged.max_col,
                ))

        self._insert_continuation_header(
            ws, insert_row, document_header_row, row_count=7
        )
        # Leave the first row below the common report header empty.  The
        # subsection title/table starts on the second row below the header.
        spacer_row = insert_row + 7
        table_insert_row = spacer_row + 1
        self._insert_rows_safely(ws, spacer_row, 4)

        # Keep images below the repeated document header aligned with the rows
        # moved by the spacer plus three-row subsection/table-header insertion.
        for image in list(ws._images):
            anchor = getattr(image, 'anchor', None)
            if not hasattr(anchor, '_from') or anchor._from.row + 1 < table_insert_row:
                continue
            anchor._from.row += 4
            if hasattr(anchor, 'to'):
                anchor.to.row += 4

        for offset, row_snapshot in enumerate(cell_snapshot):
            target_row = table_insert_row + offset
            ws.row_dimensions[target_row].height = row_heights[offset]
            for col, saved in enumerate(row_snapshot, start=1):
                cell = ws.cell(row=target_row, column=col)
                cell.value = saved['value']
                cell._style = copy.copy(saved['style'])
                cell.number_format = saved['number_format']

        for min_ro, min_col, max_ro, max_col in source_merges:
            ws.merge_cells(
                start_row=table_insert_row + min_ro, start_column=min_col,
                end_row=table_insert_row + max_ro, end_column=max_col,
            )

        # Mark the copied subsection as a continuation without depending on
        # a particular language or fixed title column.
        for col in range(1, max_col + 1):
            cell = ws.cell(row=table_insert_row, column=col)
            if str(cell.value or '').strip():
                cell.value = f"{cell.value} (계속)"
                break

        return 11

    def _insert_page13_rt_continuation(self, ws, insert_row,
                                       document_header_row, rt_header_row):
        """Create a page-13-style continuation page for overflowing RT rows.

        Rows at ``insert_row`` and below already contain the overflowing RT
        records, the RT TOTAL row, and the unused lower-page frame. Inserting
        the copied document/RT headers at that exact point completes the page
        13 frame without duplicating PAUT, MT, or PT data. The safe insertion
        also moves the original page 14 and every later page break down.
        """
        return self._insert_table_continuation(
            ws,
            insert_row,
            document_header_row,
            rt_header_row,
        )

    def _first_print_overflow_row(self, ws, start_row, row_count):
        """Return the first data row that falls onto the next printed page."""
        break_ids = sorted(int(brk.id) for brk in ws.row_breaks.brk)
        previous_breaks = [row for row in break_ids if row < start_row]
        page_start = (previous_breaks[-1] + 1) if previous_breaks else 1

        # A4 portrait printable height, adjusted by the template's print scale.
        margins = ws.page_margins
        scale = float(ws.page_setup.scale or 100) / 100.0
        usable_inches = 11.69 - float(margins.top or 0) \
            - float(margins.bottom or 0) - float(margins.footer or 0)
        capacity_points = usable_inches * 72.0 / scale
        default_height = float(ws.sheet_format.defaultRowHeight or 15)

        used_points = 0.0
        for row in range(page_start, start_row):
            used_points += float(ws.row_dimensions[row].height or default_height)

        for row in range(start_row, start_row + row_count):
            row_height = float(ws.row_dimensions[row].height or default_height)
            if used_points + row_height > capacity_points:
                return row
            used_points += row_height
        return None

    def _find_next_document_header(self, ws, after_row, search_rows=80):
        """Locate the next real monthly-report header instead of assuming an offset."""
        end_row = min(ws.max_row, after_row + search_rows)
        for row in range(after_row + 1, end_row + 1):
            for col in range(1, min(ws.max_column, 30) + 1):
                value = str(ws.cell(row=row, column=col).value or '')
                compact = value.replace(' ', '').replace('\n', '')
                if '월간용역진도보고서' in compact:
                    return row

        # Before placeholder replacement, the contract-title row may not yet
        # contain the final report wording.  The fixed 3.0 title is four rows
        # below the beginning of the document header in this template.
        for row in range(after_row + 1, end_row + 1):
            for col in range(1, min(ws.max_column, 30) + 1):
                value = str(ws.cell(row=row, column=col).value or '')
                compact = value.replace(' ', '').replace('\n', '')
                if '3.0비파괴검사현황' in compact:
                    return max(after_row + 1, row - 4)

        # Final structural fallback: template page headers begin immediately
        # after the next explicit horizontal page break.
        later_breaks = sorted(
            int(brk.id) for brk in ws.row_breaks.brk if int(brk.id) >= after_row
        )
        if later_breaks:
            return later_breaks[0] + 1
        return None

    def _renumber_ndt_status_pages(self, ws):
        """Renumber every generated page belonging to section 3.0."""
        header_rows = []
        for row in range(1, ws.max_row + 1):
            value = str(ws.cell(row=row, column=6).value or '')
            compact = value.replace(' ', '').replace('\n', '')
            if '3.0비파괴검사현황' in compact:
                header_rows.append(row)

        total_pages = len(header_rows)
        for page_number, header_row in enumerate(header_rows, start=1):
            ws.cell(row=header_row, column=16).value = (
                f"  쪽  번 호 :      {page_number}     of     {total_pages}"
            )

    def _keep_iuc_euc_headers_single_line(self, ws):
        """Keep the narrow IUC/EUC column headers on a single line."""
        import copy

        for row in ws.iter_rows():
            for cell in row:
                compact = ''.join(str(cell.value or '').split()).upper()
                if compact not in {'IUC', 'EUC'}:
                    continue

                cell.value = compact
                alignment = copy.copy(cell.alignment)
                alignment.wrap_text = False
                alignment.shrink_to_fit = True
                cell.alignment = alignment

    def _fix_ndt_section_labels(self, ws):
        """Correct known text errors retained in the report template."""
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                if 'SPETION' in cell.value:
                    cell.value = cell.value.replace('SPETION', 'SECTION')

    def _ensure_cover_cell_elements(self, ws, create_date):
        """Use cell-native cover elements that survive openpyxl save cycles."""
        import copy

        thin = Side(style='thin', color='000000')
        gray = Side(style='thin', color='BFBFBF')

        def _merge_exactly(cell_range):
            from openpyxl.worksheet.cell_range import CellRange

            target = CellRange(cell_range)
            for merged in list(ws.merged_cells.ranges):
                overlaps = not (
                    merged.max_row < target.min_row
                    or merged.min_row > target.max_row
                    or merged.max_col < target.min_col
                    or merged.min_col > target.max_col
                )
                if overlaps and str(merged) != cell_range:
                    ws.unmerge_cells(str(merged))
            if not any(str(merged) == cell_range for merged in ws.merged_cells.ranges):
                ws.merge_cells(cell_range)

        # The original diagonal was an unsupported Line shape. A diagonal-up
        # border on one merged cell produces the same slash and is preserved.
        _merge_exactly('B4:C7')
        diagonal_cell = ws['B4']
        diagonal_cell.value = None
        diagonal_cell.border = Border(
            left=thin, right=thin, top=thin, bottom=thin,
            diagonal=thin, diagonalUp=True,
        )

        # The original date was an unsupported text box. Write the creation
        # date from the export dialog into a merged cell instead.
        _merge_exactly('N4:T6')
        date_cell = ws['N4']
        date_cell.value = str(create_date or '')
        # Match the cover's revision-number typography exactly.
        date_cell.font = copy.copy(ws['N3'].font)
        date_cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=False,
        )
        date_cell.border = Border(left=gray, right=gray, top=gray, bottom=gray)

    def _populate_radiation_network_report(self, ws, target_ym):
        """Replace the section 6.0 screenshot with a real editable cell table."""
        title_row = None
        for row in range(1, ws.max_row + 1):
            row_text = ''.join(
                str(ws.cell(row=row, column=col).value or '')
                for col in range(1, min(ws.max_column, 30) + 1)
            ).replace(' ', '')
            if '1.방사선안전관리통합정보망보고자료' in row_text:
                title_row = row
                break
        if title_row is None:
            return

        table_start = title_row + 1
        next_page_start = None
        for row in range(table_start + 1, min(ws.max_row, table_start + 80) + 1):
            row_text = ''.join(
                str(ws.cell(row=row, column=col).value or '')
                for col in range(1, min(ws.max_column, 30) + 1)
            ).replace(' ', '')
            if '6.0사진대지' in row_text:
                next_page_start = max(table_start + 1, row - 4)
                break
        page_end = (next_page_start - 1) if next_page_start else table_start + 46

        # Remove only the captured report image. Logos in the common header are
        # anchored above table_start and must remain on the page.
        kept_images = []
        for image in list(ws._images):
            anchor = getattr(image, 'anchor', None)
            if hasattr(anchor, '_from'):
                image_row = anchor._from.row + 1
                if table_start <= image_row <= page_end:
                    continue
            kept_images.append(image)
        ws._images = kept_images

        source_path = os.path.join(
            os.path.expanduser('~'), 'Downloads', '발주자 보고.xlsx'
        )
        records = []
        if os.path.exists(source_path):
            try:
                source_wb = openpyxl.load_workbook(source_path, data_only=True)
                source_ws = source_wb['작업정보'] \
                    if '작업정보' in source_wb.sheetnames else source_wb.active
                target_key = str(target_ym or '').replace('-', '')[:6]

                def _text(value):
                    if value is None:
                        return ''
                    if isinstance(value, (datetime.datetime, datetime.date)):
                        return value.strftime('%Y-%m-%d')
                    return str(value).strip()

                def _clock(value):
                    text = _text(value)
                    if len(text) == 4 and text.isdigit():
                        return f'{text[:2]}:{text[2:]}'
                    return text

                for row in range(5, source_ws.max_row + 1):
                    year_month = _text(source_ws.cell(row, 1).value).replace('-', '')[:6]
                    if target_key and year_month != target_key:
                        continue
                    workers = [
                        _text(source_ws.cell(row, col).value)
                        for col in range(10, 15)
                    ]
                    workers = ', '.join(value for value in workers if value)

                    source_name = ''
                    source_activity = ''
                    source_count = ''
                    for name, activity_col, count_col in (
                        ('Ir-192', 21, 22), ('Co-60', 23, 24),
                        ('Se-75', 25, 26), ('Am-241', 27, 28),
                    ):
                        activity = source_ws.cell(row, activity_col).value
                        count = source_ws.cell(row, count_col).value
                        if activity not in (None, '') or count not in (None, ''):
                            source_name = name
                            source_activity = activity if activity is not None else ''
                            source_count = count if count is not None else ''
                            break

                    generator = ' / '.join(
                        value for value in (
                            _text(source_ws.cell(row, 29).value),
                            _text(source_ws.cell(row, 30).value),
                        ) if value
                    )
                    records.append([
                        year_month,
                        _text(source_ws.cell(row, 2).value),
                        _text(source_ws.cell(row, 3).value),
                        _text(source_ws.cell(row, 4).value),
                        _text(source_ws.cell(row, 5).value),
                        source_ws.cell(row, 6).value or '',
                        _text(source_ws.cell(row, 7).value),
                        _text(source_ws.cell(row, 8).value),
                        _text(source_ws.cell(row, 9).value),
                        workers,
                        _clock(source_ws.cell(row, 15).value),
                        _clock(source_ws.cell(row, 16).value),
                        source_ws.cell(row, 17).value or '',
                        _text(source_ws.cell(row, 18).value),
                        _text(source_ws.cell(row, 19).value),
                        source_ws.cell(row, 20).value or '',
                        source_name,
                        source_activity,
                        source_count,
                        generator,
                        source_ws.cell(row, 31).value or '',
                        _text(source_ws.cell(row, 32).value),
                    ])
                source_wb.close()
            except Exception as exc:
                print(f'[WARN] 발주자 보고자료를 읽지 못했습니다: {exc}')

        # Protect the following photo page when a month contains more rows than
        # the original screenshot area. The page break and print area move with
        # the inserted rows; drawings need the same explicit anchor adjustment.
        planned_data_start = table_start + 2
        available_rows = max(1, page_end - planned_data_start + 1)
        extra_rows = max(0, len(records) - available_rows)
        if extra_rows and next_page_start:
            self._insert_rows_safely(ws, next_page_start, extra_rows)
            for image in list(ws._images):
                anchor = getattr(image, 'anchor', None)
                if not hasattr(anchor, '_from') or anchor._from.row + 1 < next_page_start:
                    continue
                anchor._from.row += extra_rows
                if hasattr(anchor, 'to'):
                    anchor.to.row += extra_rows
            page_end += extra_rows

        # Clear the old screenshot area while retaining the surrounding page.
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row >= table_start and merged.max_row <= page_end:
                ws.unmerge_cells(str(merged))
        for row in range(table_start, page_end + 1):
            for col in range(2, 24):
                ws.cell(row=row, column=col).value = None

        group_row = table_start
        header_row = table_start + 1
        data_start = table_start + 2
        groups = (
            (2, 7, '작업장 정보'),
            (8, 14, '작업 기본정보'),
            (15, 17, '검사대상물'),
            (18, 20, '최대사용선원'),
            (21, 22, '방사선발생장치'),
            (23, 23, '비고'),
        )
        headers = (
            '년월', '업체코드', '검사업체명', '작업장코드', '작업장명',
            '일일최대량', '작업일', '작업장소', '작업조장', '작업자',
            '시작', '종료', '작업수량', '이름', '재질', '최대두께',
            '선원', '방사능량', '수량', '최대전압/전류', '발생장치수량', '비고',
        )
        dark_fill = PatternFill('solid', fgColor='BFBFBF')
        orange_fill = PatternFill('solid', fgColor='F4B183')
        yellow_fill = PatternFill('solid', fgColor='FFF200')
        header_font = Font(name='맑은 고딕', size=7, bold=True)
        body_font = Font(name='맑은 고딕', size=7)
        center = Alignment(
            horizontal='center', vertical='center', wrap_text=True,
            shrink_to_fit=True,
        )

        for start_col, end_col, label in groups:
            if start_col < end_col:
                ws.merge_cells(
                    start_row=group_row, start_column=start_col,
                    end_row=group_row, end_column=end_col,
                )
            cell = ws.cell(group_row, start_col, label)
            cell.fill = dark_fill
            cell.font = header_font
            cell.alignment = center
        for offset, label in enumerate(headers, start=2):
            cell = ws.cell(header_row, offset, label)
            cell.fill = orange_fill if offset <= 7 else yellow_fill
            cell.font = header_font
            cell.alignment = center

        ws.row_dimensions[group_row].height = 20
        ws.row_dimensions[header_row].height = 34
        if not records:
            ws.merge_cells(
                start_row=data_start, start_column=2,
                end_row=data_start, end_column=23,
            )
            ws.cell(data_start, 2).value = (
                f'{target_ym} 발주자 보고자료 없음'
            )
            ws.cell(data_start, 2).alignment = center
            ws.cell(data_start, 2).font = body_font
            records_end = data_start
        else:
            for row_offset, record in enumerate(records):
                target_row = data_start + row_offset
                for col_offset, value in enumerate(record, start=2):
                    cell = ws.cell(target_row, col_offset, value)
                    cell.font = body_font
                    cell.alignment = center
                ws.row_dimensions[target_row].height = 24
            records_end = data_start + len(records) - 1

        for row in range(group_row, records_end + 1):
            for col in range(2, 24):
                ws.cell(row, col).border = self.border_thin

    def _populate_owner_report_rows(self, ws, target_ym, source_path):
        """Fill the existing section 6.0 template rows from a selected file."""
        if not source_path:
            return
        if not os.path.exists(source_path):
            raise FileNotFoundError(f'발주자 보고 파일을 찾을 수 없습니다: {source_path}')

        title_row = None
        for row in range(1, ws.max_row + 1):
            row_text = ''.join(
                str(ws.cell(row=row, column=col).value or '')
                for col in range(1, min(ws.max_column, 30) + 1)
            ).replace(' ', '')
            if '1.방사선안전관리통합정보망보고자료' in row_text:
                title_row = row
                break
        if title_row is None:
            raise ValueError('템플릿에서 6.0 발주자 보고자료 표를 찾을 수 없습니다.')

        # The edited V70 template uses rows 630~632 for its three-level header.
        data_start = title_row + 4
        data_end = data_start + 42
        for row in range(data_start, data_end + 1):
            for col in range(2, 24):
                cell = ws.cell(row=row, column=col)
                if type(cell).__name__ != 'MergedCell':
                    cell.value = None

        source_wb = openpyxl.load_workbook(source_path, data_only=True)
        try:
            if '작업정보' not in source_wb.sheetnames:
                raise ValueError("발주자 보고 파일에 '작업정보' 시트가 없습니다.")
            source_ws = source_wb['작업정보']
            target_key = str(target_ym or '').replace('-', '')[:6]

            def _text(value):
                if value is None:
                    return ''
                if isinstance(value, (datetime.datetime, datetime.date)):
                    return value.strftime('%Y-%m-%d')
                return str(value).strip()

            def _clock(value):
                text = _text(value)
                digits = re.sub(r'\D', '', text)
                if len(digits) == 3:
                    digits = f'0{digits}'
                if len(digits) == 4:
                    return f'{digits[:2]}:{digits[2:]}'
                return text

            records = []
            for row in range(5, source_ws.max_row + 1):
                year_month = _text(source_ws.cell(row, 1).value)
                year_month = re.sub(r'\D', '', year_month)[:6]
                if target_key and year_month != target_key:
                    continue
                record = [
                    year_month,
                    _text(source_ws.cell(row, 2).value),
                    _text(source_ws.cell(row, 3).value),
                    _text(source_ws.cell(row, 4).value),
                    _text(source_ws.cell(row, 5).value),
                    source_ws.cell(row, 6).value or '',
                    _text(source_ws.cell(row, 7).value),
                    _text(source_ws.cell(row, 8).value),
                    _text(source_ws.cell(row, 9).value),
                    *[
                        _text(source_ws.cell(row, col).value)
                        for col in range(10, 15)
                    ],
                    _clock(source_ws.cell(row, 15).value),
                    _clock(source_ws.cell(row, 16).value),
                    source_ws.cell(row, 17).value or '',
                    _text(source_ws.cell(row, 18).value),
                    _text(source_ws.cell(row, 19).value),
                    source_ws.cell(row, 20).value or '',
                    source_ws.cell(row, 25).value or '',
                    source_ws.cell(row, 26).value or '',
                ]
                records.append(record)
        finally:
            source_wb.close()

        if len(records) > data_end - data_start + 1:
            raise ValueError(
                f'발주자 보고자료가 {len(records)}행으로 16페이지 입력 가능 '
                f'행수({data_end - data_start + 1}행)를 초과했습니다.'
            )

        body_font = Font(name='맑은 고딕', size=6)
        body_alignment = Alignment(
            horizontal='center', vertical='center',
            wrap_text=True, shrink_to_fit=True,
        )
        for row_offset, record in enumerate(records):
            target_row = data_start + row_offset
            for column, value in enumerate(record, start=2):
                cell = ws.cell(row=target_row, column=column)
                cell.value = value
                cell.font = body_font
                cell.alignment = body_alignment

    def _trim_trailing_blank_print_pages(self, ws):
        """Stop printing at the end of the last page containing real content.

        Dynamic row insertion expands the template print area and shifts every
        manual row break.  That can leave a fully blank page after the last
        report page.  Preserve the complete last content page by trimming at
        its following page boundary, rather than at the last populated row.
        """
        from openpyxl.utils import get_column_letter

        print_area_text = str(ws.print_area or '')
        area_match = re.search(
            r"\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)",
            print_area_text.upper(),
        )
        if not area_match:
            return

        min_col_name, min_row_text, max_col_name, max_row_text = area_match.groups()
        from openpyxl.utils.cell import column_index_from_string
        min_col = column_index_from_string(min_col_name)
        max_col = column_index_from_string(max_col_name)
        min_row = int(min_row_text)
        print_max_row = int(max_row_text)

        last_content_row = min_row
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=min(print_max_row, ws.max_row),
            min_col=min_col,
            max_col=max_col,
        ):
            if any(cell.value not in (None, '') for cell in row):
                last_content_row = row[0].row

        # Images are printable content even though their cells have no value.
        for image in list(ws._images):
            anchor = getattr(image, 'anchor', None)
            if not hasattr(anchor, '_from'):
                continue
            image_bottom = anchor._from.row + 1
            if hasattr(anchor, 'to'):
                image_bottom = max(image_bottom, anchor.to.row + 1)
            last_content_row = max(last_content_row, image_bottom)

        later_breaks = sorted(
            int(brk.id)
            for brk in ws.row_breaks.brk
            if last_content_row <= int(brk.id) < print_max_row
        )
        if not later_breaks:
            return

        last_page_end = later_breaks[0]
        ws.print_area = (
            f"${get_column_letter(min_col)}${min_row}:"
            f"${get_column_letter(max_col)}${last_page_end}"
        )
        ws.row_breaks.brk = [
            brk for brk in ws.row_breaks.brk
            if int(brk.id) < last_page_end
        ]

    def _rt_reject_defects(self, row):
        """Return reject-grade RT defects encoded as ``grade/defect``."""
        method = str(row.get('검사방법', '') or '').strip().upper()
        if method != 'RT':
            return []

        defects = []
        section_info = str(row.get('구간정보', '') or '')
        for item in re.split(r'[,;\n]+', section_info):
            match = re.match(r'^\s*([1-4])\s*/\s*([A-Z]+)\s*$', item.upper())
            if not match:
                continue
            grade = int(match.group(1))
            if grade >= 3:
                defects.append(match.group(2))
        return defects

    def _rt_film_count(self, row):
        """Count populated RT section-info slots (one slot equals one film)."""
        if str(row.get('검사방법', '') or '').strip().upper() != 'RT':
            return 0
        section_info = str(row.get('구간정보', '') or '')
        return sum(
            1 for item in re.split(r'[,;\n]+', section_info)
            if item.strip()
        )

    def _write_welder_defect_summary(
        self, ws, defects_by_welder, retests_by_welder,
        original_counts_by_welder, section_number, section_keywords,
    ):
        """Write welder defects and A/B quantities into a 4.0 summary table."""
        import copy

        if not defects_by_welder and not retests_by_welder and not original_counts_by_welder:
            return

        title_row = None
        for row in range(1, ws.max_row + 1):
            row_text = ''.join(
                str(ws.cell(row=row, column=col).value or '')
                for col in range(1, ws.max_column + 1)
            ).replace(' ', '').replace('\n', '').upper()
            if (
                section_number in row_text
                and ('불량률' in row_text or '불량율' in row_text)
                and any(keyword.upper() in row_text for keyword in section_keywords)
            ):
                title_row = row
                break
        if title_row is None:
            return

        header_row = None
        id_col = None
        name_col = None
        header_cols = {}
        for row in range(title_row + 1, min(title_row + 7, ws.max_row + 1)):
            current_headers = {}
            current_id_col = None
            for col in range(1, ws.max_column + 1):
                label = ''.join(str(ws.cell(row=row, column=col).value or '').split()).upper()
                if label == 'ID':
                    current_id_col = col
                elif label:
                    current_headers.setdefault(label, col)
            if current_id_col is not None:
                header_row = row
                id_col = current_id_col
                header_cols = current_headers
                name_col = current_headers.get('성명')
                break
        if header_row is None:
            return

        retest_col = None
        original_count_col = None
        defect_rate_col = None
        for row in range(title_row + 1, header_row + 1):
            for col in range(1, ws.max_column + 1):
                label = ''.join(str(ws.cell(row=row, column=col).value or '').split()).upper()
                if '불합격부' in label and '재검사' in label and ('(A)' in label or label.endswith('A')):
                    retest_col = col
                if '당초' in label and '시공물량' in label and ('(B)' in label or label.endswith('B')):
                    original_count_col = col
                if '용접' in label and ('불량률' in label or '불량율' in label) and 'A/B' in label:
                    defect_rate_col = col
            if (
                retest_col is not None
                and original_count_col is not None
                and defect_rate_col is not None
            ):
                break

        total_row = None
        for row in range(header_row + 1, min(header_row + 30, ws.max_row + 1)):
            values = [
                ''.join(str(ws.cell(row=row, column=col).value or '').split()).upper()
                for col in range(1, ws.max_column + 1)
            ]
            if any(value in {'TOTAL', '총계'} for value in values):
                total_row = row
                break
        if total_row is None:
            return

        data_start = header_row + 1
        available_rows = total_row - data_start
        welders = sorted(
            set(defects_by_welder)
            | set(retests_by_welder)
            | set(original_counts_by_welder)
        )
        if len(welders) > available_rows:
            extra_rows = len(welders) - available_rows
            style_source_row = max(data_start, total_row - 1)
            row_style = [
                copy.copy(ws.cell(row=style_source_row, column=col)._style)
                for col in range(1, ws.max_column + 1)
            ]
            row_height = ws.row_dimensions[style_source_row].height
            self._insert_rows_safely(ws, total_row, extra_rows)
            for row in range(total_row, total_row + extra_rows):
                ws.row_dimensions[row].height = row_height
                for col, style in enumerate(row_style, start=1):
                    ws.cell(row=row, column=col)._style = copy.copy(style)
            total_row += extra_rows

        defect_columns = {
            code: header_cols.get(code)
            for code in {code for counts in defects_by_welder.values() for code in counts}
            if header_cols.get(code) is not None
        }
        for offset, welder in enumerate(welders):
            row = data_start + offset
            id_cell = ws.cell(row=row, column=id_col)
            id_cell.value = welder
            id_alignment = copy.copy(id_cell.alignment)
            id_alignment.wrap_text = False
            id_alignment.shrink_to_fit = True
            id_cell.alignment = id_alignment
            if name_col is not None:
                name_cell = ws.cell(row=row, column=name_col)
                name_cell.value = self.WELDER_NAMES.get(welder.upper(), '')
                name_alignment = copy.copy(name_cell.alignment)
                name_alignment.wrap_text = False
                name_alignment.shrink_to_fit = True
                name_cell.alignment = name_alignment
            for code, count in defects_by_welder.get(welder, {}).items():
                col = defect_columns.get(code)
                if col is not None:
                    ws.cell(row=row, column=col).value = count
            if retest_col is not None:
                ws.cell(row=row, column=retest_col).value = retests_by_welder.get(welder, 0)
            if original_count_col is not None:
                ws.cell(row=row, column=original_count_col).value = (
                    original_counts_by_welder.get(welder, 0)
                )
            if defect_rate_col is not None:
                retest_qty = retests_by_welder.get(welder, 0)
                original_qty = original_counts_by_welder.get(welder, 0)
                rate_cell = ws.cell(row=row, column=defect_rate_col)
                rate_cell.value = retest_qty / original_qty if original_qty else None
                rate_cell.number_format = '0.00%'
                rate_alignment = copy.copy(rate_cell.alignment)
                rate_alignment.wrap_text = False
                rate_alignment.shrink_to_fit = True
                rate_cell.alignment = rate_alignment

        # Recalculate each defect total without disturbing the other totals.
        for col in set(defect_columns.values()):
            total = sum(
                self._safe_float(ws.cell(row=row, column=col).value)
                for row in range(data_start, data_start + len(welders))
            )
            ws.cell(row=total_row, column=col).value = int(total)
        if retest_col is not None:
            ws.cell(row=total_row, column=retest_col).value = sum(
                retests_by_welder.values()
            )
        if original_count_col is not None:
            ws.cell(row=total_row, column=original_count_col).value = sum(
                original_counts_by_welder.values()
            )
        if defect_rate_col is not None:
            total_retest = sum(retests_by_welder.values())
            total_original = sum(original_counts_by_welder.values())
            total_rate_cell = ws.cell(row=total_row, column=defect_rate_col)
            total_rate_cell.value = (
                total_retest / total_original if total_original else None
            )
            total_rate_cell.number_format = '0.00%'
            total_rate_alignment = copy.copy(total_rate_cell.alignment)
            total_rate_alignment.wrap_text = False
            total_rate_alignment.shrink_to_fit = True
            total_rate_cell.alignment = total_rate_alignment

    def _populate_process_photo_pages(self, ws, process_photos):
        """Place registered process photos into the template's 7.0 pages."""
        from openpyxl.drawing.image import Image as XLImage

        base_dir = os.path.dirname(os.path.abspath(self._history_path))
        valid_photos = []
        for photo in process_photos:
            process = str(photo.get('process', '')).strip().upper()
            if process not in {'PAUT', 'MT', 'RT', 'PT'}:
                continue
            stored_path = str(photo.get('file_path', '') or '')
            image_path = (
                stored_path if os.path.isabs(stored_path)
                else os.path.abspath(os.path.join(base_dir, stored_path))
            )
            if os.path.isfile(image_path):
                saved_photo = dict(photo)
                saved_photo['_resolved_path'] = image_path
                valid_photos.append(saved_photo)
        process_photos = valid_photos

        original_photo_starts = []
        for row in range(600, ws.max_row + 1):
            row_text = ''.join(
                str(ws.cell(row=row, column=col).value or '')
                for col in range(1, ws.max_column + 1)
            ).replace(' ', '').upper()
            if 'SEC.16' in row_text and any(
                keyword in row_text
                for keyword in ('PAUT', '위상배열초음파탐상검사', '자분탐상검사', '방사선투과검사')
            ):
                original_photo_starts.append(row - 9)
        photo_section_start = min(original_photo_starts) if original_photo_starts else None

        pt_photos = [
            photo for photo in process_photos
            if str(photo.get('process', '')).upper() == 'PT'
        ]
        pt_title_row = self._ensure_pt_photo_page(ws) if pt_photos else None

        processes_with_photos = {
            str(photo.get('process', '')).upper()
            for photo in process_photos
            if str(photo.get('process', '')).strip()
        }
        self._remove_empty_process_photo_pages(ws, processes_with_photos)

        page_layouts = {}
        title_keywords = {
            'PAUT': ('위상배열초음파탐상검사', 'PAUT'),
            'MT': ('자분탐상검사',),
            'RT': ('방사선투과검사',),
        }
        for row in range(1, ws.max_row + 1):
            row_text = ''.join(
                str(ws.cell(row=row, column=col).value or '')
                for col in range(1, ws.max_column + 1)
            ).replace(' ', '').upper()
            for process, keywords in title_keywords.items():
                if process in page_layouts:
                    continue
                if (
                    any(keyword.upper() in row_text for keyword in keywords)
                    and 'SEC.16' in row_text
                    and row > 600
                ):
                    page_layouts[process] = {
                        'title_row': row,
                        'image_row': row + 2,
                        'caption_row': row + 42,
                    }
        # PT may have moved when empty PAUT/MT/RT pages were deleted, so locate
        # every surviving page again instead of relying on its old row number.
        if pt_title_row is not None:
            for row in range(600, ws.max_row + 1):
                row_text = ''.join(
                    str(ws.cell(row=row, column=col).value or '')
                    for col in range(1, ws.max_column + 1)
                ).replace(' ', '').upper()
                if '침투탐상검사' in row_text and 'SEC.16' in row_text:
                    page_layouts['PT'] = {
                        'title_row': row,
                        'image_row': row + 2,
                        'caption_row': row + 42,
                    }
                    break
        descriptions = {
            'PAUT': '위상배열초음파탐상검사',
            'MT': '자분탐상검사',
            'RT': '방사선투과검사',
            'PT': '침투탐상검사',
        }
        # Add as many continuation pages as needed (four photos per page).
        for process, layout in sorted(
            page_layouts.items(), key=lambda item: item[1]['title_row'], reverse=True
        ):
            photo_count = sum(
                1 for photo in process_photos
                if str(photo.get('process', '')).upper() == process
            )
            extra_pages = max(0, (photo_count - 1) // 4)
            source_title_row = layout['title_row']
            for continuation_no in range(2, extra_pages + 2):
                source_title_row = self._clone_process_photo_page(
                    ws, source_title_row, continuation_no
                )

        # All lower pages may have moved. Rebuild a list of every surviving
        # base/continuation page for each process.
        layouts_by_process = {process: [] for process in descriptions}
        scan_keywords = dict(title_keywords)
        scan_keywords['PT'] = ('침투탐상검사',)
        for row in range(600, ws.max_row + 1):
            row_text = ''.join(
                str(ws.cell(row=row, column=col).value or '')
                for col in range(1, ws.max_column + 1)
            ).replace(' ', '').upper()
            if 'SEC.16' not in row_text:
                continue
            for process, keywords in scan_keywords.items():
                if any(keyword.upper() in row_text for keyword in keywords):
                    layouts_by_process[process].append({
                        'title_row': row,
                        'image_row': row + 2,
                    })
                    break

        for process, process_layouts in layouts_by_process.items():
            photos = sorted(
                [
                    photo for photo in process_photos
                    if str(photo.get('process', '')).upper() == process
                ],
                key=lambda photo: (
                    str(photo.get('date', '')), str(photo.get('joint_no', ''))
                ),
            )
            for page_index, layout in enumerate(process_layouts):
                page_photos = photos[page_index * 4:(page_index + 1) * 4]
                if not page_photos:
                    continue

                photo_slots = (
                    ('C', layout['image_row'], 3, layout['image_row'] + 18),
                    ('M', layout['image_row'], 13, layout['image_row'] + 18),
                    ('C', layout['image_row'] + 21, 3, layout['image_row'] + 39),
                    ('M', layout['image_row'] + 21, 13, layout['image_row'] + 39),
                )
                for index, photo in enumerate(page_photos):
                    image = XLImage(photo['_resolved_path'])
                    max_width, max_height = 310, 180
                    scale = min(
                        max_width / image.width,
                        max_height / image.height,
                        1.0,
                    )
                    image.width *= scale
                    image.height *= scale
                    anchor_col, anchor_row, caption_col, caption_row = photo_slots[index]
                    image.anchor = f"{anchor_col}{anchor_row}"
                    ws.add_image(image)

                    caption = ' / '.join(filter(None, [
                        str(photo.get('date', '')),
                        str(photo.get('location', '')),
                        str(photo.get('joint_no', '')),
                        str(photo.get('description', '')),
                    ]))
                    caption_end_col = 12 if caption_col == 3 else 23
                    for merged in list(ws.merged_cells.ranges):
                        if (
                            merged.min_row <= caption_row <= merged.max_row
                            and merged.min_col <= caption_end_col
                            and merged.max_col >= caption_col
                        ):
                            ws.unmerge_cells(str(merged))
                    ws.merge_cells(
                        start_row=caption_row, start_column=caption_col,
                        end_row=caption_row, end_column=caption_end_col,
                    )
                    caption_cell = ws.cell(row=caption_row, column=caption_col)
                    caption_cell.value = caption
                    caption_cell.alignment = Alignment(
                        horizontal='center', vertical='center',
                        wrap_text=True, shrink_to_fit=True,
                    )

        # The template contains additional legacy photo-log pages after the
        # three SEC.16 pages.  Print only the generated process-photo pages;
        # when no valid photo exists, stop at the preceding report page.
        surviving_titles = [
            layout['title_row']
            for layouts in layouts_by_process.values()
            for layout in layouts
        ]
        if photo_section_start is not None:
            print_end_row = (
                max(surviving_titles) + 44
                if surviving_titles else photo_section_start - 1
            )
            self._set_print_area_end(ws, print_end_row)

    def _set_print_area_end(self, ws, end_row):
        """Set the vertical print limit and discard later manual breaks."""
        from openpyxl.worksheet.cell_range import CellRange

        if not ws.print_area:
            return
        try:
            print_range = CellRange(str(ws.print_area).replace('$', ''))
            print_range.max_row = max(print_range.min_row, int(end_row))
            ws.print_area = str(print_range)
            ws.row_breaks.brk = [
                brk for brk in ws.row_breaks.brk
                if int(brk.id) < print_range.max_row
            ]
        except Exception:
            pass

    def _remove_empty_process_photo_pages(self, ws, processes_with_photos):
        """Delete complete template pages for processes without photos."""
        from openpyxl.worksheet.cell_range import CellRange

        keywords = {
            'PAUT': ('위상배열초음파탐상검사', 'PAUT'),
            'MT': ('자분탐상검사',),
            'RT': ('방사선투과검사',),
            'PT': ('침투탐상검사',),
        }
        pages_to_delete = []
        for row in range(600, ws.max_row + 1):
            row_text = ''.join(
                str(ws.cell(row=row, column=col).value or '')
                for col in range(1, ws.max_column + 1)
            ).replace(' ', '').upper()
            if 'SEC.16' not in row_text:
                continue
            for process, process_keywords in keywords.items():
                if process in processes_with_photos:
                    continue
                if any(keyword.upper() in row_text for keyword in process_keywords):
                    # Photo-page titles sit nine rows below the manual page
                    # boundary (PAUT 676:729, MT 730:783, RT 784:837).
                    pages_to_delete.append((row - 9, process))
                    break

        page_rows = 54
        for page_start, _process in sorted(pages_to_delete, reverse=True):
            page_end = page_start + page_rows - 1

            # Remove drawings owned by the deleted page and shift every later
            # logo/photo upward with the worksheet content.
            kept_images = []
            for image in list(ws._images):
                anchor = getattr(image, 'anchor', None)
                if not hasattr(anchor, '_from'):
                    kept_images.append(image)
                    continue
                image_row = anchor._from.row + 1
                if page_start <= image_row <= page_end:
                    continue
                if image_row > page_end:
                    anchor._from.row -= page_rows
                    if hasattr(anchor, 'to'):
                        anchor.to.row -= page_rows
                kept_images.append(image)
            ws._images = kept_images

            self._delete_rows_safely(ws, page_start, page_rows)

            if ws.print_area:
                try:
                    print_range = CellRange(str(ws.print_area).replace('$', ''))
                    if print_range.max_row > page_end:
                        print_range.max_row -= page_rows
                    elif print_range.max_row >= page_start:
                        print_range.max_row = page_start - 1
                    ws.print_area = str(print_range)
                except Exception:
                    pass

    def _clone_process_photo_page(self, ws, source_title_row, continuation_no):
        """Clone one 54-row photo page directly after its source page."""
        import copy
        from io import BytesIO
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.worksheet.pagebreak import Break

        page_rows = 54
        source_start = source_title_row - 9
        source_end = source_start + page_rows - 1
        insert_row = source_end + 1
        max_col = ws.max_column
        cell_snapshot = [
            [
                (
                    ws.cell(row=row, column=col).value,
                    copy.copy(ws.cell(row=row, column=col)._style),
                    copy.copy(ws.cell(row=row, column=col).alignment),
                    ws.cell(row=row, column=col).number_format,
                )
                for col in range(1, max_col + 1)
            ]
            for row in range(source_start, source_end + 1)
        ]
        row_snapshot = [
            (ws.row_dimensions[row].height, ws.row_dimensions[row].hidden)
            for row in range(source_start, source_end + 1)
        ]
        merge_snapshot = [
            (
                merged.min_row - source_start, merged.min_col,
                merged.max_row - source_start, merged.max_col,
            )
            for merged in list(ws.merged_cells.ranges)
            if merged.min_row >= source_start and merged.max_row <= source_end
        ]
        image_snapshot = []
        for image in list(ws._images):
            anchor = getattr(image, 'anchor', None)
            if hasattr(anchor, '_from') and source_start <= anchor._from.row + 1 <= source_end:
                image_snapshot.append((image, copy.deepcopy(anchor)))

        self._insert_rows_safely(ws, insert_row, page_rows)
        for existing_image in list(ws._images):
            anchor = getattr(existing_image, 'anchor', None)
            if hasattr(anchor, '_from') and anchor._from.row + 1 >= insert_row:
                anchor._from.row += page_rows
                if hasattr(anchor, 'to'):
                    anchor.to.row += page_rows

        for offset, saved_row in enumerate(cell_snapshot):
            target_row = insert_row + offset
            height, hidden = row_snapshot[offset]
            ws.row_dimensions[target_row].height = height
            ws.row_dimensions[target_row].hidden = hidden
            for col, (value, style, alignment, number_format) in enumerate(saved_row, 1):
                cell = ws.cell(row=target_row, column=col)
                cell.value = value
                cell._style = copy.copy(style)
                cell.alignment = copy.copy(alignment)
                cell.number_format = number_format
        for min_row, min_col, max_row, max_col in merge_snapshot:
            ws.merge_cells(
                start_row=insert_row + min_row, start_column=min_col,
                end_row=insert_row + max_row, end_column=max_col,
            )

        row_shift = insert_row - source_start
        for source_image, saved_anchor in image_snapshot:
            buffer = BytesIO()
            image_ref = source_image.ref
            if hasattr(image_ref, 'copy') and hasattr(image_ref, 'save'):
                image_ref.copy().save(
                    buffer, format=getattr(image_ref, 'format', None) or 'PNG'
                )
            elif isinstance(image_ref, (str, os.PathLike)):
                with open(image_ref, 'rb') as stream:
                    buffer.write(stream.read())
            else:
                old_position = image_ref.tell() if hasattr(image_ref, 'tell') else None
                if hasattr(image_ref, 'seek'):
                    image_ref.seek(0)
                buffer.write(image_ref.read())
                if old_position is not None and hasattr(image_ref, 'seek'):
                    image_ref.seek(old_position)
            buffer.seek(0)
            new_image = XLImage(buffer)
            new_image.width = source_image.width
            new_image.height = source_image.height
            new_image._photo_continuation_stream = buffer
            new_image.anchor = copy.deepcopy(saved_anchor)
            new_image.anchor._from.row += row_shift
            if hasattr(new_image.anchor, 'to'):
                new_image.anchor.to.row += row_shift
            ws.add_image(new_image)

        new_title_row = insert_row + 9
        title_cell = ws.cell(row=new_title_row, column=3)
        base_title = re.sub(
            r'\s*\(계속\s*\d+\)\s*$', '', str(title_cell.value or '').strip()
        )
        title_cell.value = f"{base_title} (계속 {continuation_no})"
        page_end = insert_row + page_rows - 1
        if not any(int(brk.id) == page_end for brk in ws.row_breaks.brk):
            ws.row_breaks.append(Break(id=page_end))
        return new_title_row

    def _ensure_pt_photo_page(self, ws):
        """Clone a template photo page after RT when PT photos exist."""
        import copy
        from io import BytesIO
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.worksheet.pagebreak import Break

        # Reuse an existing generated/template PT page when present.
        for row in range(600, ws.max_row + 1):
            text = ''.join(
                str(ws.cell(row=row, column=col).value or '')
                for col in range(1, ws.max_column + 1)
            ).replace(' ', '').upper()
            if '침투탐상검사' in text and ('SEC.16' in text or 'PT' in text):
                return row

        source_title_row = None
        rt_title_row = None
        for row in range(600, ws.max_row + 1):
            text = ''.join(
                str(ws.cell(row=row, column=col).value or '')
                for col in range(1, ws.max_column + 1)
            ).replace(' ', '').upper()
            if source_title_row is None and '자분탐상검사' in text and 'SEC.16' in text:
                source_title_row = row
            if rt_title_row is None and '방사선투과검사' in text and 'SEC.16' in text:
                rt_title_row = row
        if source_title_row is None or rt_title_row is None:
            return None

        page_rows = 54
        source_start = source_title_row - 9
        source_end = source_start + page_rows - 1
        insert_row = rt_title_row - 9 + page_rows
        max_col = ws.max_column

        cells = []
        for row in range(source_start, source_end + 1):
            cells.append([
                (
                    ws.cell(row=row, column=col).value,
                    copy.copy(ws.cell(row=row, column=col)._style),
                    copy.copy(ws.cell(row=row, column=col).alignment),
                    ws.cell(row=row, column=col).number_format,
                )
                for col in range(1, max_col + 1)
            ])
        row_dimensions = [
            (
                ws.row_dimensions[row].height,
                ws.row_dimensions[row].hidden,
            )
            for row in range(source_start, source_end + 1)
        ]
        merges = [
            (
                merged.min_row - source_start, merged.min_col,
                merged.max_row - source_start, merged.max_col,
            )
            for merged in list(ws.merged_cells.ranges)
            if merged.min_row >= source_start and merged.max_row <= source_end
        ]
        source_images = []
        for image in list(ws._images):
            anchor = getattr(image, 'anchor', None)
            if not hasattr(anchor, '_from'):
                continue
            image_row = anchor._from.row + 1
            if source_start <= image_row <= source_end:
                source_images.append((image, copy.deepcopy(anchor)))

        self._insert_rows_safely(ws, insert_row, page_rows)
        # openpyxl does not move drawings when rows are inserted.  Keep every
        # existing logo/photo below the new PT page aligned with its content.
        for existing_image in list(ws._images):
            anchor = getattr(existing_image, 'anchor', None)
            if not hasattr(anchor, '_from') or anchor._from.row + 1 < insert_row:
                continue
            anchor._from.row += page_rows
            if hasattr(anchor, 'to'):
                anchor.to.row += page_rows
        for offset, row_cells in enumerate(cells):
            target_row = insert_row + offset
            height, hidden = row_dimensions[offset]
            ws.row_dimensions[target_row].height = height
            ws.row_dimensions[target_row].hidden = hidden
            for col, (value, style, alignment, number_format) in enumerate(
                row_cells, start=1
            ):
                cell = ws.cell(row=target_row, column=col)
                cell.value = value
                cell._style = copy.copy(style)
                cell.alignment = copy.copy(alignment)
                cell.number_format = number_format
        for min_row, min_col, max_row, max_col in merges:
            ws.merge_cells(
                start_row=insert_row + min_row, start_column=min_col,
                end_row=insert_row + max_row, end_column=max_col,
            )

        row_shift = insert_row - source_start
        for source_image, saved_anchor in source_images:
            image_buffer = BytesIO()
            image_ref = source_image.ref
            if hasattr(image_ref, 'copy') and hasattr(image_ref, 'save'):
                image_ref.copy().save(
                    image_buffer,
                    format=getattr(image_ref, 'format', None) or 'PNG',
                )
            elif isinstance(image_ref, (str, os.PathLike)):
                with open(image_ref, 'rb') as stream:
                    image_buffer.write(stream.read())
            else:
                old_position = image_ref.tell() if hasattr(image_ref, 'tell') else None
                if hasattr(image_ref, 'seek'):
                    image_ref.seek(0)
                image_buffer.write(image_ref.read())
                if old_position is not None and hasattr(image_ref, 'seek'):
                    image_ref.seek(old_position)
            image_buffer.seek(0)
            new_image = XLImage(image_buffer)
            new_image.width = source_image.width
            new_image.height = source_image.height
            new_image._process_photo_page_stream = image_buffer
            new_image.anchor = copy.deepcopy(saved_anchor)
            new_image.anchor._from.row += row_shift
            if hasattr(new_image.anchor, 'to'):
                new_image.anchor.to.row += row_shift
            ws.add_image(new_image)

        # Keep RT and the generated PT page as separate printed pages.
        rt_page_end = insert_row - 1
        pt_page_end = insert_row + page_rows - 1
        existing_breaks = {int(brk.id) for brk in ws.row_breaks.brk}
        if rt_page_end not in existing_breaks:
            ws.row_breaks.append(Break(id=rt_page_end))
        if pt_page_end not in existing_breaks:
            ws.row_breaks.append(Break(id=pt_page_end))

        pt_title_row = insert_row + (source_title_row - source_start)
        ws.cell(row=pt_title_row, column=3).value = '1.4 침투탐상검사(PT)(SEC.16)'
        return pt_title_row

    def generate_report(self, history_path, target_ym, output_path, doc_num="01",
                        create_date=None, owner_report_path=None):

        self._history_path = history_path

        if create_date is None:
            from datetime import datetime
            create_date = str(datetime.today().date())
        """
        history_path: daily_work_history.json 경로
        target_ym: "YYYY-MM" 형식 (예: "2026-08")
        """
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template not found at {self.template_path}")
            
        if not os.path.exists(history_path):
            raise FileNotFoundError(f"History file not found at {history_path}")

        with open(history_path, 'r', encoding='utf-8') as f:
            history_data = json.load(f)

        # 1. 데이터 필터링 및 집계
        # The selected year-month is the inclusive end of a cumulative report.
        # Include every valid history date from the beginning through that
        # month's final day (for example, 2027-08 includes through 2027-08-31).
        target_dates = sorted(
            d for d in history_data.keys()
            if re.match(r'^\d{4}-\d{2}-\d{2}$', str(d))
            and str(d)[:7] <= target_ym
        )
        if not target_dates:
            print(f"No data found through {target_ym}")
            return None

        # 집계 구조 (1.2 요약용: 업체+구간+라인+관경+규격별 그룹)
        qty_summary = {}
        ndt_groups = {
            'PAUT': defaultdict(lambda: {'count': 0, 'qty': 0.0}),
            'MT': defaultdict(lambda: {'count': 0, 'qty': 0.0}),
            'RT': defaultdict(lambda: {'count': 0, 'qty': 0.0}),
            'PT': defaultdict(lambda: {'count': 0, 'qty': 0.0})
        }
        # 세부 현황용 (2.x 전체 레코드 - 날짜순)
        ndt_details = {'PAUT': [], 'MT': [], 'RT': [], 'PT': []}
        process_photos = []
        # 4.0/1.1 RT 용접사별 불합격 결함 수
        rt_defects_by_welder = defaultdict(lambda: defaultdict(int))
        rt_retests_by_welder = defaultdict(int)
        rt_original_counts_by_welder = defaultdict(int)
        paut_retest_lengths_by_welder = defaultdict(float)
        paut_original_lengths_by_welder = defaultdict(float)

        # 물량표(qty) 처리를 위해 시작일의 전일누계와 전체 금일작업 합산을 구함
        first_day_data = history_data[target_dates[0]]
        last_day_data = history_data[target_dates[-1]]

        for key, val in first_day_data.get('qty_data', {}).items():
            qty_summary[key] = {
                '예상량': val.get('예상량', ''),
                '전월누계': val.get('전일누계', '0'), # 월간 보고서이므로 첫날의 전일누계가 전월누계
                '금월작업': 0.0,
                '총누계': val.get('총누계', '0'),
                '공정률': val.get('공정률', ''),
                '불량': 0,
                '불량률': ''
            }

        # 날짜별 금일작업 누적
        for d in target_dates:
            day_data = history_data[d]
            process_photos.extend(day_data.get('process_photos', []))
            for key, val in day_data.get('qty_data', {}).items():
                if key not in qty_summary:
                    qty_summary[key] = {'예상량': val.get('예상량', ''), '전월누계': '0', '금월작업': 0.0, '총누계': '0', '공정률': '', '불량': 0, '불량률': ''}
                
                try:
                    today_work = float(str(val.get('금일작업', '0')).replace(',', '') or 0)
                    qty_summary[key]['금월작업'] += today_work
                except ValueError:
                    pass
                
                # 불량(결함) 수량도 누적 (숫자일 경우)
                try:
                    defect = float(str(val.get('불량', '0')).replace(',', '') or 0)
                    qty_summary[key]['불량'] += defect
                except ValueError:
                    pass

            # NDT 결과 집계
            for row in day_data.get('ndt_results', []):
                method = str(row.get('검사방법', '')).strip().upper()
                if not method or method not in ndt_groups:
                    continue

                if method == 'RT':
                    welder = str(row.get('용접사', '') or '').strip()
                    if welder:
                        for defect_code in self._rt_reject_defects(row):
                            rt_defects_by_welder[welder][defect_code] += 1
                        joint_no = str(row.get('Joint No.', '') or '').strip().upper()
                        film_count = self._rt_film_count(row)
                        if joint_no.endswith('R1'):
                            rt_retests_by_welder[welder] += film_count
                        elif joint_no:
                            rt_original_counts_by_welder[welder] += film_count
                elif method == 'PAUT':
                    welder = str(row.get('용접사', '') or '').strip()
                    joint_no = str(row.get('Joint No.', '') or '').strip().upper()
                    inspection_length = self._safe_float(row.get('PAUT'))
                    if welder and inspection_length > 0:
                        if joint_no.endswith('R1'):
                            paut_retest_lengths_by_welder[welder] += inspection_length
                        elif joint_no:
                            paut_original_lengths_by_welder[welder] += inspection_length
                
                company = str(row.get('업체', '')).strip()
                if not company: continue # 빈 행 무시
                
                section = str(row.get('구간', '')).strip()
                line_no = str(row.get('라인번호', '')).strip()
                size = str(row.get('관경', '')).strip()
                spec = str(row.get('규격', '')).strip()
                
                # 수량 (길이 또는 매수)
                qty_val = 0.0
                if method == 'PAUT': qty_val = self._safe_float(row.get('PAUT'))
                elif method == 'MT': qty_val = self._safe_float(row.get('MT'))
                elif method == 'PT': qty_val = self._safe_float(row.get('PT'))
                elif method == 'RT':
                    # RT는 OR + RE 매수 합산
                    or_val = self._safe_float(row.get('RT_OR'))
                    re_val = self._safe_float(row.get('RT_RE'))
                    qty_val = or_val + re_val
                
                group_key = (company, section, line_no, size, spec)
                ndt_groups[method][group_key]['count'] += 1
                ndt_groups[method][group_key]['qty'] += qty_val

                # 2.x 세부현황용: 개별 레코드 저장 (날짜 포함)
                if method == 'RT':
                    d_ori = self._safe_float(row.get('RT_OR'))
                    d_re  = self._safe_float(row.get('RT_RE'))
                else:
                    d_ori = qty_val
                    d_re  = 0.0
                if d_ori + d_re > 0:
                    ndt_details[method].append({
                        'key': group_key,
                        'ori': d_ori, 're': d_re, 'qty': d_ori + d_re
                    })

        # 마지막 날의 총누계와 공정률을 최종 반영
        for key, val in last_day_data.get('qty_data', {}).items():
            if key in qty_summary:
                qty_summary[key]['총누계'] = val.get('총누계', '0')
                qty_summary[key]['공정률'] = val.get('공정률', '')
                if val.get('불량률'):
                    qty_summary[key]['불량률'] = val.get('불량률', '')

        # 2. 엑셀 쓰기 준비
        wb = openpyxl.load_workbook(self.template_path)
        
        # 엑셀 열 때 '명명된 범위' 복구 창이 뜨는 것을 방지하기 위해 불필요한 이름 정의 삭제
        if hasattr(wb, 'defined_names'):
            try:
                wb.defined_names.clear()
            except:
                pass
                
        # 외부 수식 참조(externalLink) 찌꺼기로 인한 복구 창 방지
        if hasattr(wb, '_external_links'):
            wb._external_links = []
            
        # [표지] 시트의 병합 셀이 openpyxl 저장 시 풀리는 버그 방지를 위한 백업
        ws = wb.active

        # The template already has the page 6/7 boundary at row 229. Insert
        # the missing page 5/6 boundary at row 192 without replacing row 229
        # or any later boundary (including the page 7/8 break at row 278).
        from openpyxl.worksheet.pagebreak import Break
        template_breaks = sorted(ws.row_breaks.brk, key=lambda brk: int(brk.id))
        if not any(int(brk.id) == 192 for brk in template_breaks):
            template_breaks.append(Break(id=192))
        ws.row_breaks.brk = sorted(
            template_breaks, key=lambda brk: int(brk.id)
        )

        # Capture the template's compact body-row height before dynamic inserts
        # move row 466. Header rows keep their original heights.
        detail_body_row_height = (
            ws.row_dimensions[466].height
            or ws.sheet_format.defaultRowHeight
            or 15
        )




        # 역순으로 채워야 행 삽입 시 위에 있는 인덱스(row number)가 변하지 않음
        # 따라서 아래에서부터 표를 처리합니다.
        
        # 표 위치들을 리스트로 묶어서 역순 처리
        # is_detail=True: 2.x 세부(전체 레코드), False: 1.x 요약(그룹)
        sections = [
            ('pt_2', 'PT', True), ('rt_2', 'RT', True), ('mt_2', 'MT', True), ('paut_2', 'PAUT', True),
            ('pt_1', 'PT', False), ('rt_1', 'RT', False), ('mt_1', 'MT', False), ('paut_1', 'PAUT', False)
        ]
        section_item_counts = {}
        # MT/PT가 없고 RT만 있는 경우 13페이지의 PT 하단 빈 행을
        # RT 확장 공간으로 우선 재사용한다.
        rt_uses_page13_blanks = (
            bool(ndt_details['RT'])
            and not ndt_details['MT']
            and not ndt_details['PT']
        )
        
        for section_key, method, is_detail in sections:
            header_row = self.table_markers[section_key]
            start_row = header_row + 2
            
            # TOTAL 행을 찾음 (header_row 이후 빈 행이 아닌 "TOTAL"이 있는 행)
            total_row = start_row
            for r in range(start_row, start_row + 20):
                val = str(ws.cell(row=r, column=2).value or "").strip()
                if "TOTAL" in val.upper() or "총계" in val:
                    total_row = r
                    break
            
            # 데이터 삽입: 2.x는 전체 레코드, 1.x는 요약 그룹
            if is_detail:
                data_items = [
                    (rec['key'], {'count': 1, 'qty': rec['qty'], 'ori': rec['ori'], 're': rec['re']})
                    for rec in ndt_details[method]
                ]
            else:
                data_items = []
                current_company = None
                sub_count = 0
                sub_qty = 0.0
                
                # 업체명 기준으로 1차 정렬
                sorted_groups = sorted(ndt_groups[method].items(), key=lambda x: str(x[0][0]))
                
                for k, v in sorted_groups:
                    comp = k[0]
                    if current_company is not None and comp != current_company:
                        data_items.append(
                            ((f"[{current_company} 소계]", "", "", "", ""), {'count': sub_count, 'qty': sub_qty, 'ori': sub_qty, 're': 0.0, 'is_subtotal': True})
                        )
                        sub_count = 0
                        sub_qty = 0.0
                        
                    current_company = comp
                    data_items.append(
                        (k, {'count': v['count'], 'qty': v['qty'], 'ori': v['qty'], 're': 0.0, 'is_subtotal': False})
                    )
                    sub_count += v['count']
                    sub_qty += v['qty']
                    
                if current_company is not None:
                    data_items.append(
                        ((f"[{current_company} 소계]", "", "", "", ""), {'count': sub_count, 'qty': sub_qty, 'ori': sub_qty, 're': 0.0, 'is_subtotal': True})
                    )
            num_items = len(data_items)
            section_item_counts[section_key] = num_items
            
            def _write_safe(r, c, val):
                cell = ws.cell(row=r, column=c)
                if type(cell).__name__ == 'MergedCell':
                    new_cell = Cell(ws, row=r, column=c)
                    ws._cells[(r, c)] = new_cell
                    cell = new_cell
                cell.value = val

            def safe_merge(sr, sc, er, ec):
                from openpyxl.utils import get_column_letter
                coord = f"{get_column_letter(sc)}{sr}:{get_column_letter(ec)}{er}"
                overlaps = []
                for m in list(ws.merged_cells.ranges):
                    # Check for overlap
                    if m.bounds[0] <= ec and m.bounds[2] >= sc and m.bounds[1] <= er and m.bounds[3] >= sr:
                        if str(m) == coord:
                            return # Already perfectly merged
                        overlaps.append(m)
                
                # Unmerge any overlapping regions first to prevent Excel file corruption
                for m in overlaps:
                    try:
                        ws.unmerge_cells(str(m))
                    except:
                        pass
                
                try: ws.merge_cells(coord)
                except: pass
            
            # 기존에 2개의 빈 행이 있다고 가정하고(TOTAL 위에), 모자라면 삽입
            existing_empty_rows = total_row - start_row
            
            if num_items > existing_empty_rows:
                rows_to_insert = num_items - existing_empty_rows
                # 안전한 행 삽입: 아래 시트(13페이지 등) 병합 셀 구조 보존
                self._insert_rows_safely(ws, start_row, rows_to_insert)
                total_row += rows_to_insert
                # 모든 기존 마커 업데이트 (삽입된 행 아래에 있는 마커들만 += rows_to_insert)
                for k, v in self.table_markers.items():
                    if v >= start_row:
                        self.table_markers[k] += rows_to_insert

                cleaned = 0
                if section_key == 'rt_2' and rt_uses_page13_blanks:
                    # RT 삽입으로 아래로 밀린 2.4 PT의 TOTAL 아래부터
                    # 현재 수동 페이지 구분선까지 빈 행을 제거한다. 삽입/삭제
                    # 수가 상쇄되므로 기존 14페이지 내용은 움직이지 않는다.
                    pt_marker = self.table_markers['pt_2']
                    later_breaks = sorted(
                        int(brk.id) for brk in ws.row_breaks.brk
                        if int(brk.id) >= pt_marker
                    )
                    page_end = later_breaks[0] if later_breaks else None
                    pt_total_row = None
                    if page_end is not None:
                        for candidate in range(pt_marker + 2, page_end + 1):
                            if 'TOTAL' in str(
                                ws.cell(row=candidate, column=2).value or ''
                            ).upper():
                                pt_total_row = candidate
                                break

                    scan_row = pt_total_row + 1 if pt_total_row else None
                    while (
                        scan_row is not None
                        and page_end is not None
                        and cleaned < rows_to_insert
                        and scan_row <= page_end
                    ):
                        is_empty = all(
                            not str(ws.cell(row=scan_row, column=c).value or '').strip()
                            for c in range(1, 25)
                        )
                        if not is_empty:
                            scan_row += 1
                            continue
                        self._delete_rows_safely(ws, scan_row, 1)
                        for k, v in self.table_markers.items():
                            if v >= scan_row:
                                self.table_markers[k] -= 1
                        page_end -= 1
                        cleaned += 1

                # 일반 섹션은 기존처럼 TOTAL 직후의 연속 빈 행을 정리한다.
                if not (section_key == 'rt_2' and rt_uses_page13_blanks):
                    scan_row = total_row + 1
                    while cleaned < rows_to_insert:
                        is_empty = all(
                            not str(ws.cell(row=scan_row, column=c).value or '').strip()
                            for c in range(1, 25)
                        )
                        if is_empty:
                            self._delete_rows_safely(ws, scan_row, 1)
                            total_row -= 1
                            for k, v in self.table_markers.items():
                                if v >= scan_row:
                                    self.table_markers[k] -= 1
                            cleaned += 1
                        else:
                            break
            else:
                rows_to_insert = 0

            current_row = start_row
            seq_num = 1
            for idx, ((comp, sec, line, size, spec), vals) in enumerate(data_items):
                count = vals['count']
                qty = vals['qty']
                unit = "매" if method == "RT" else "m"
                is_subtotal = vals.get('is_subtotal', False)
                if is_subtotal:
                    unit = ""

                # Clear reusable template values before writing this detail row.
                for clear_col in range(17, 24):
                    clear_cell = ws.cell(row=current_row, column=clear_col)
                    if type(clear_cell).__name__ != 'MergedCell':
                        clear_cell.value = None
                
                # 셀 위치: 2:업체, 3:순번, 4:Section, 6:Line No., 10:관경, 12:용접개소, 14:규격, 16:단위, 17:길이
                ws.cell(row=current_row, column=2).value = comp
                if is_subtotal:
                    ws.cell(row=current_row, column=3).value = ""
                else:
                    ws.cell(row=current_row, column=3).value = str(seq_num)
                    seq_num += 1
                ws.cell(row=current_row, column=4).value = sec
                ws.cell(row=current_row, column=6).value = line
                ws.cell(row=current_row, column=10).value = size
                ws.cell(row=current_row, column=12).value = count
                ws.cell(row=current_row, column=14).value = spec
                ws.cell(row=current_row, column=16).value = unit
                
                ws.cell(row=current_row, column=16).value = unit

                ori_v = vals.get('ori', qty)
                re_v  = vals.get('re', 0.0)
                tot_v = ori_v + re_v

                def _fmt(v):
                    return f"{v:.4f}" if v % 1 != 0 else str(int(v))

                _write_safe(current_row, 17, _fmt(ori_v) if ori_v > 0 else '')
                if re_v > 0:
                    _write_safe(current_row, 18, _fmt(re_v))  # RE' 앵커(col18)
                # PAUT는 항상 TOTAL, 세부현황도 TOTAL 표시
                if method == 'PAUT' or is_detail:
                    _write_safe(current_row, 20, _fmt(tot_v) if tot_v > 0 else '')
                
                # 병합 처리 (템플릿 양식에 맞춤)
                safe_merge(current_row, 4, current_row, 5) # Section
                safe_merge(current_row, 6, current_row, 9) # Line No.
                safe_merge(current_row, 10, current_row, 11) # 관경
                safe_merge(current_row, 12, current_row, 13) # 용접개소
                safe_merge(current_row, 14, current_row, 15) # 규격
                
                # PAUT/MT/RT/PT 모두 동일: RE'(col18~19)만 병합, ORI'(17)·TOTAL(20)은 개별 셀
                safe_merge(current_row, 18, current_row, 19) # RE'
                safe_merge(current_row, 20, current_row, 21) # TOTAL
                safe_merge(current_row, 22, current_row, 23) # 비고
                    
                # Line No. 글씨 다 보이도록 높이 조절
                if current_row >= 401:
                    ws.row_dimensions[current_row].height = detail_body_row_height

                # 폰트, 정렬 적용 (병합된 칸 전체에 정렬 속성을 먹여야 엑셀이 줄바꿈을 정상 인식함)
                for c_idx in range(2, 24):
                    c_cell = ws.cell(row=current_row, column=c_idx)
                    if is_subtotal:
                        import copy
                        bold_font = copy.copy(self.font_normal)
                        bold_font.bold = True
                        c_cell.font = bold_font
                        
                        align = copy.copy(self.align_center)
                        if c_idx == 2:
                            align.wrap_text = False
                            align.shrink_to_fit = True
                        c_cell.alignment = align
                    else:
                        c_cell.font = self.font_normal
                        c_cell.alignment = self.align_center
                    c_cell.border = self.border_thin

                # Keep long Line No. values on one line inside the widened F:I area.
                ws.cell(row=current_row, column=6).alignment = Alignment(
                    horizontal='center', vertical='center',
                    wrap_text=False, shrink_to_fit=True
                )

                current_row += 1

            # Dynamic row operations can move the TOTAL row; locate it again.
            for candidate_row in range(current_row, current_row + 40):
                candidate_value = str(
                    ws.cell(row=candidate_row, column=2).value or ''
                ).strip().upper()
                if 'TOTAL' in candidate_value:
                    total_row = candidate_row
                    break

            if num_items == 0:
                # Keep the method visible while clearly distinguishing "no
                # activity" from a missing or failed data load.
                safe_merge(start_row, 2, start_row, 23)
                no_data_cell = ws.cell(row=start_row, column=2)
                no_data_cell.value = '해당 월 검사실적 없음'
                no_data_cell.font = self.font_normal
                no_data_cell.alignment = self.align_center
                no_data_cell.border = self.border_thin
                if start_row >= 401:
                    ws.row_dimensions[start_row].height = detail_body_row_height
                current_row = start_row + 1
                
            # 빈 행(데이터 없는 행)들 스타일 정리 및 PAUT 병합 처리 준비
            for r in range(current_row, total_row):
                if r >= 401:
                    ws.row_dimensions[r].height = detail_body_row_height
                # 빈 행도 기본적으로 다른 칸들도 병합을 유지해야 템플릿 양식이 안 깨짐
                safe_merge(r, 4, r, 5)
                safe_merge(r, 6, r, 9)
                safe_merge(r, 10, r, 11)
                safe_merge(r, 12, r, 13)
                safe_merge(r, 14, r, 15)
                
                # 빈 행도 RE'(col18~19)만 병합 (ORI'·TOTAL 개별 유지)
                safe_merge(r, 18, r, 19)
                safe_merge(r, 20, r, 21) # TOTAL
                safe_merge(r, 22, r, 23) # 비고
                for c_idx in range(2, 24):
                    cell = ws.cell(row=r, column=c_idx)
                    cell.border = self.border_thin

            # TOTAL 행 값 쓰기 및 스타일 보정
            total_ori = sum(vals.get('ori', vals['qty']) for _, vals in data_items if not vals.get('is_subtotal', False))
            total_re = sum(vals.get('re', 0.0) for _, vals in data_items if not vals.get('is_subtotal', False))
            total_qty = total_ori + total_re
            
            # TOTAL 행의 17, 18, 19, 20 칸 배경색(파란색)을 동일하게 맞춤 (첫 번째 셀 서식 복사)
            fill_style = ws.cell(row=total_row, column=2).fill
            if method == 'PAUT':
                import copy
                for c_idx in range(17, 21):
                    c = ws.cell(row=total_row, column=c_idx)
                    if fill_style:
                        c.fill = copy.copy(fill_style)

            safe_merge(total_row, 22, total_row, 23) # 비고
            if total_row >= 401:
                ws.row_dimensions[total_row].height = detail_body_row_height

            if num_items == 0:
                _write_safe(total_row, 17, 0)
                _write_safe(total_row, 20, 0)
                ws.cell(row=total_row, column=17).alignment = self.align_center
                ws.cell(row=total_row, column=20).alignment = self.align_center

            if total_qty > 0:
                from openpyxl.styles import Alignment as _Align
                shrink_align = _Align(horizontal='center', vertical='center', shrink_to_fit=True)
                if method == 'PAUT':
                    safe_merge(total_row, 20, total_row, 21)
                    _write_safe(total_row, 17, f"{total_ori:.4f}" if total_ori % 1 != 0 else str(int(total_ori)))
                    ws.cell(row=total_row, column=17).alignment = shrink_align
                    if total_re > 0:
                        _write_safe(total_row, 18, f"{total_re:.4f}" if total_re % 1 != 0 else str(int(total_re)))
                        ws.cell(row=total_row, column=18).alignment = shrink_align
                    _write_safe(total_row, 20, f"{total_qty:.4f}" if total_qty % 1 != 0 else str(int(total_qty)))
                    ws.cell(row=total_row, column=20).alignment = shrink_align
                else:
                    _write_safe(total_row, 17, f"{total_qty:.4f}" if total_qty % 1 != 0 else str(int(total_qty)))
                    ws.cell(row=total_row, column=17).alignment = shrink_align
                    if is_detail:
                        # Detail tables display ORI' and TOTAL in separate
                        # columns, so repeat the combined quantity in TOTAL.
                        safe_merge(total_row, 20, total_row, 21)
                        _write_safe(
                            total_row,
                            20,
                            f"{total_qty:.4f}"
                            if total_qty % 1 != 0 else str(int(total_qty)),
                        )
                        ws.cell(row=total_row, column=20).alignment = shrink_align

            # When either a 1.2 summary table or a 2.x detail table overflows,
            # continue it on a real new page. Repeat both the seven-row document
            # header and the subsection/table header.  Repeat as many times as
            # necessary; following MT/RT/PT/4.0 content moves down as a block.
            remaining_start = start_row
            remaining_count = num_items
            previous_section_breaks = sorted(
                int(brk.id)
                for brk in ws.row_breaks.brk
                if int(brk.id) < header_row
            )
            document_header_row = (
                previous_section_breaks[-1] + 1
                if previous_section_breaks else 1
            )
            while remaining_count > 0:
                continuation_row = self._first_print_overflow_row(
                    ws, remaining_start, remaining_count
                )
                if continuation_row is None:
                    break

                rows_on_current_page = continuation_row - remaining_start
                if section_key == 'rt_2' and rt_uses_page13_blanks:
                    inserted_rows = self._insert_page13_rt_continuation(
                        ws, continuation_row, document_header_row, header_row
                    )
                else:
                    inserted_rows = self._insert_table_continuation(
                        ws, continuation_row, document_header_row, header_row
                    )
                from openpyxl.worksheet.pagebreak import Break
                if not any(
                    int(brk.id) == continuation_row - 1
                    for brk in ws.row_breaks.brk
                ):
                    ws.row_breaks.append(Break(id=continuation_row - 1))

                for key, marker_row in self.table_markers.items():
                    if marker_row >= continuation_row:
                        self.table_markers[key] += inserted_rows

                remaining_count -= rows_on_current_page
                remaining_start = continuation_row + inserted_rows

        # When the three method-detail tables on this page have no records,
        # distribute all unused vertical space across MT->RT, RT->PT, and
        # PT->page-bottom.  Including the final gap prevents the lower half of
        # the page from remaining empty while the tables crowd the top.
        if all(section_item_counts.get(key, 0) == 0 for key in ('mt_2', 'rt_2', 'pt_2')):
            method_keys = ('mt_2', 'rt_2', 'pt_2')
            total_rows = []
            for index, current_key in enumerate(method_keys):
                current_marker = self.table_markers[current_key]
                search_end = (
                    self.table_markers[method_keys[index + 1]]
                    if index + 1 < len(method_keys)
                    else ws.max_row + 1
                )
                current_total = None
                for row in range(current_marker + 2, search_end):
                    if 'TOTAL' in str(ws.cell(row=row, column=2).value or '').upper():
                        current_total = row
                        break
                total_rows.append(current_total)

            later_breaks = sorted(
                int(brk.id) for brk in ws.row_breaks.brk
                if int(brk.id) >= self.table_markers['pt_2']
            )
            page_end = later_breaks[0] if later_breaks else None

            if all(row is not None for row in total_rows) and page_end is not None:
                gap_limits = (
                    (total_rows[0] + 1, self.table_markers['rt_2']),
                    (total_rows[1] + 1, self.table_markers['pt_2']),
                    (total_rows[2] + 1, page_end + 1),
                )
                default_height = ws.sheet_format.defaultRowHeight or 15
                gap_rows = []
                for gap_start, gap_end in gap_limits:
                    blank_spacer_rows = []
                    for spacer_row in range(gap_start, gap_end):
                        is_blank = all(
                            not str(ws.cell(row=spacer_row, column=col).value or '').strip()
                            for col in range(1, 24)
                        )
                        if is_blank:
                            blank_spacer_rows.append(spacer_row)
                    gap_rows.append(blank_spacer_rows)

                # Equalize the number of spacer rows instead of assigning
                # different row heights to each gap.  This keeps every visible
                # grid cell the same height while preserving equal spacing.
                total_spacers = sum(len(rows) for rows in gap_rows)
                base_count, remainder = divmod(total_spacers, 3)
                desired_counts = [
                    base_count + (1 if index < remainder else 0)
                    for index in range(3)
                ]

                first_delta = desired_counts[0] - len(gap_rows[0])
                rt_marker = self.table_markers['rt_2']
                rt_title_row = rt_marker - 1
                if first_delta > 0:
                    self._insert_rows_safely(ws, rt_title_row, first_delta)
                elif first_delta < 0:
                    self._delete_rows_safely(
                        ws, rt_title_row + first_delta, -first_delta
                    )
                if first_delta:
                    self.table_markers['rt_2'] += first_delta
                    self.table_markers['pt_2'] += first_delta
                    total_rows[1] += first_delta
                    total_rows[2] += first_delta
                    page_end += first_delta

                second_delta = desired_counts[1] - len(gap_rows[1])
                pt_marker = self.table_markers['pt_2']
                pt_title_row = pt_marker - 1
                if second_delta > 0:
                    self._insert_rows_safely(ws, pt_title_row, second_delta)
                elif second_delta < 0:
                    self._delete_rows_safely(
                        ws, pt_title_row + second_delta, -second_delta
                    )
                if second_delta:
                    self.table_markers['pt_2'] += second_delta
                    total_rows[2] += second_delta
                    page_end += second_delta

                third_delta = desired_counts[2] - len(gap_rows[2])
                if third_delta > 0:
                    self._insert_rows_safely(ws, page_end, third_delta)
                    page_end += third_delta
                elif third_delta < 0:
                    self._delete_rows_safely(
                        ws, total_rows[2] + 1, -third_delta
                    )
                    page_end += third_delta

                uniform_gaps = (
                    (total_rows[0] + 1, self.table_markers['rt_2'] - 1),
                    (total_rows[1] + 1, self.table_markers['pt_2'] - 1),
                    (total_rows[2] + 1, page_end + 1),
                )
                for gap_start, gap_end in uniform_gaps:
                    for spacer_row in range(gap_start, gap_end):
                        ws.row_dimensions[spacer_row].height = default_height

        # 마지막으로 맨 위의 1.1 물량표 처리
        qty_start_row = self.table_markers['qty'] + 1
        qty_mapping = {
            'PAUT_300A이상': qty_start_row,
            'PAUT_300A이상-야간': qty_start_row + 1,
            'PAUT_250A': qty_start_row + 2,
            'PAUT_200A': qty_start_row + 3,
            'PAUT_200A-야간': qty_start_row + 4,
            'PAUT_소계': qty_start_row + 5,
            
            'RT_150A~100A': qty_start_row + 6,
            'RT_150A~100A-야간': qty_start_row + 7,
            'RT_80A이하': qty_start_row + 8,
            'RT_80A이하-야간': qty_start_row + 9,
            'RT_소계': qty_start_row + 10,
            
            'MT_전체(주간)': qty_start_row + 11,
            'MT_전체(야간)': qty_start_row + 12,
            
            'PT_전체(주간)': qty_start_row + 13,
            'PT_전체(야간)': qty_start_row + 14,
        }

        for key, r_idx in qty_mapping.items():
            if key in qty_summary:
                data = qty_summary[key]
                # 컬럼: 2:방법, 4:규격, 7:예상량, 9:전월누계, 11:금월작업, 13:총누계, 15:공정률, 17:불량, 19:불량률
                ws.cell(row=r_idx, column=7).value = self._format_num(data['예상량'])
                ws.cell(row=r_idx, column=9).value = self._format_num(data['전월누계'])
                ws.cell(row=r_idx, column=11).value = self._format_num(data['금월작업'])
                ws.cell(row=r_idx, column=13).value = self._format_num(data['총누계'])
                ws.cell(row=r_idx, column=15).value = data['공정률']
                ws.cell(row=r_idx, column=17).value = self._format_num(data['불량'])
                ws.cell(row=r_idx, column=19).value = data['불량률']

        # 사용자 요청: 관경/규격 축소, 단위/비고 확장
        ws.column_dimensions['C'].width = 4.0   # 순번
        ws.column_dimensions['J'].width = 5.0   # 관경 (8.0 → 5.0)
        ws.column_dimensions['N'].width = 2.5   # 규격 (3.5 → 2.5)
        ws.column_dimensions['O'].width = 8.0   # 단위 (5.6 → 8.0)
        ws.column_dimensions['P'].width = 8.0   # 비고 (5.0 → 8.0)
        # ORI(Q)=9, RE(R+S)=4.5+4.5=9, TOTAL(T+U)=4.5+4.5=9 → 동일 너비
        # Rebalance widths so Line No. fits on one line without widening the page.
        for col in ('F', 'G', 'H', 'I'):
            ws.column_dimensions[col].width = 6.5
        for col in ('J', 'K'):
            ws.column_dimensions[col].width = 3.5
        for col in ('L', 'M'):
            ws.column_dimensions[col].width = 3.0
        for col in ('N', 'O'):
            ws.column_dimensions[col].width = 3.5
        ws.column_dimensions['P'].width = 5.0

        ws.column_dimensions['Q'].width = 9.0
        ws.column_dimensions['R'].width = 4.5
        ws.column_dimensions['S'].width = 4.5
        ws.column_dimensions['T'].width = 4.5
        ws.column_dimensions['U'].width = 4.5
        ws.column_dimensions['V'].width = 0.0   # 비고 우측 끝 숨김
        ws.column_dimensions['V'].width = 2.25
        ws.column_dimensions['V'].hidden = False
        ws.column_dimensions['W'].width = 0.0   # 비고 우측 끝 숨김
        ws.column_dimensions['W'].width = 2.25
        ws.column_dimensions['W'].hidden = False

        # 사용자 요청: 모든 표의 좌측(2열), 우측(21열) 외곽선을 굵은 실선으로 복원/강제 설정
        medium = Side(style='medium')
        def apply_outer_borders(start_r, end_r):
            for r in range(start_r, end_r + 1):
                lc = ws.cell(row=r, column=2)
                lc.border = Border(left=medium, right=lc.border.right, top=lc.border.top, bottom=lc.border.bottom)
                rc = ws.cell(row=r, column=23)
                rc.border = Border(left=rc.border.left, right=medium, top=rc.border.top, bottom=rc.border.bottom)
                
        # 1. 상단 물량표 외곽선 교정
        if self.table_markers['qty'] > 0:
            apply_outer_borders(self.table_markers['qty'] + 1, self.table_markers['qty'] + 14)
            
        # 2. 하단 데이터 표 외곽선 교정 (헤더 제외, 데이터 행부터 TOTAL 전까지)
        for section_key in ['paut_1', 'paut_2', 'rt_1', 'rt_2', 'mt_1', 'mt_2', 'pt_1', 'pt_2']:
            if section_key not in self.table_markers: continue
            
            s_row = self.table_markers[section_key]
            if s_row > 0:
                e_row = s_row
                while True:
                    val = str(ws.cell(row=e_row, column=2).value).strip().upper()
                    if 'TOTAL' in val or '계' == val or e_row > s_row + 200:
                        break
                    e_row += 1
                apply_outer_borders(s_row, e_row - 1)

        # 문자열 치환 (문서번호, 연월, 날짜, 지사명 등)
        y, m = target_ym.split('-')
        import datetime
        replacements = {
            '[[보고서_연월]]': f"{y}년 {int(m)}월",
            '[[보고서_월]]': f"{int(m)}월",
            '[[문서번호]]': doc_num,
            '[[작성일자]]': create_date,
            '[[계약명]]': "2026년 중앙지사 열수송관 비파괴검사용역 단가계약",
            '[[지사명]]': "중앙지사",
            '2025년 동탄지사 열수송관 비파괴검사용역 단가계약': "2026년 중앙지사 열수송관 비파괴검사용역 단가계약",
            '2025년 동탄지사 열배관  비파괴검사용역 단가계약': "2026년 중앙지사 열수송관 비파괴검사용역 단가계약",
            '2025년 동탄지사 열배관 비파괴검사용역 단가계약': "2026년 중앙지사 열수송관 비파괴검사용역 단가계약",
            '동 탄 지 사': "중 앙 지 사",
            '분 당 사 업 소': "중 앙 지 사"
        }
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        val_str = str(cell.value)
                        
                        if isinstance(cell.value, str):
                            for k, v in replacements.items():
                                if k in val_str:
                                    cell.value = val_str.replace(k, str(v))
                                    val_str = str(cell.value)
                                    
                        # 타이틀 텍스트 무조건 포맷 강제 적용 (모든 페이지의 제목 칸 타겟)
                        val_no_spaces = val_str.replace(' ', '').replace('\xa0', '').replace('\n', '')
                        
                        is_title = False
                        if ('단가계약' in val_no_spaces or '계약명' in val_no_spaces) and ('월간용역' in val_no_spaces or '보고서' in val_no_spaces):
                            is_title = True
                        elif cell.column == 6 and isinstance(cell.value, str) and val_no_spaces.startswith('='):
                            # 수식으로 제목을 참조하는 경우 (예: =F41, =F$41, =$F$41 등)
                            match = re.search(r'\$?F\$?(\d+)', val_no_spaces, re.IGNORECASE)
                            if match:
                                ref_row = int(match.group(1))
                                # 참조된 셀의 실제 값이 제목 키워드를 포함하는지 확인
                                ref_val = str(sheet.cell(row=ref_row, column=6).value or '').replace(' ', '').replace('\xa0', '').replace('\n', '')
                                if ('단가계약' in ref_val or '계약명' in ref_val) and ('월간용역' in ref_val or '보고서' in ref_val):
                                    is_title = True
                                # 참조 셀에 값이 없는 경우 행 번호 패턴으로 폴백 (41, 81, 121... 등 40행 간격)
                                elif ref_row % 40 == 1:
                                    is_title = True
                                    
                        if is_title:
                            contract_name = replacements.get('[[계약명]]', '2026년 중앙지사 열수송관 비파괴검사용역 단가계약')
                            # 무조건 괄호와 줄바꿈이 포함된 포맷으로 강제 덮어쓰기 (수식도 문자열로 덮어씌움)
                            cell.value = f"【{contract_name}】\n월 간 용 역 진 도 보 고 서"
                            
                            # Alignment and Font are imported at module level.
                            import copy
                            
                            title_font = Font(name='맑은 고딕', size=9, bold=True)
                            title_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.font = title_font
                            cell.alignment = title_align
                            
                            for r_idx in range(cell.row, cell.row + 4):
                                for c_idx in range(1, 30):
                                    c = sheet.cell(row=r_idx, column=c_idx)
                                    if c.alignment:
                                        new_align = copy.copy(c.alignment)
                                        new_align.wrap_text = True
                                        c.alignment = new_align
                                    else:
                                        c.alignment = Alignment(wrap_text=True)


        # [표지] 시트 병합 셀 강제 복구 (openpyxl 버그 방지)
        # Dynamic continuation pages change the section total, so recalculate
        # the 3.0 page labels after all row and page operations are complete.
        self._write_welder_defect_summary(
            ws, rt_defects_by_welder, rt_retests_by_welder,
            rt_original_counts_by_welder, '1.1', ('방사선투과검사', 'RT'),
        )
        self._write_welder_defect_summary(
            ws, {}, paut_retest_lengths_by_welder,
            paut_original_lengths_by_welder, '1.2',
            ('위상배열초음파탐상검사', 'PAUT'),
        )
        self._populate_owner_report_rows(ws, target_ym, owner_report_path)
        self._populate_process_photo_pages(ws, process_photos)
        self._fix_ndt_section_labels(ws)
        self._ensure_cover_cell_elements(ws, create_date)
        self._renumber_ndt_status_pages(ws)
        self._keep_iuc_euc_headers_single_line(ws)

        # Preserve the visible right edge of merged template boxes and headers.
        for sheet in wb.worksheets:
            self._repair_merged_right_borders(sheet)

        # Prevent blank right-side pages caused by the widened V:W remarks area.
        # Keep vertical pagination automatic while fitting the report to one
        # printed page horizontally.
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.scale = None

        # Hide the original trailing spacer rows only while they are still
        # blank. Dynamic continuation pages may move real content into these
        # coordinates, which must never be hidden.
        for blank_row in range(570, 573):
            is_blank = all(
                not str(ws.cell(row=blank_row, column=col).value or '').strip()
                for col in range(1, min(ws.max_column, 23) + 1)
            )
            if is_blank:
                ws.row_dimensions[blank_row].height = 0
                ws.row_dimensions[blank_row].hidden = True

        self._trim_trailing_blank_print_pages(ws)

        wb.save(output_path)
        return output_path

    def _safe_float(self, val):
        try:
            return float(str(val).replace(',', ''))
        except (ValueError, TypeError):
            return 0.0
            
    def _format_num(self, val):
        try:
            f = float(str(val).replace(',', ''))
            if f == 0: return "-"
            if f % 1 == 0: return int(f)
            return f
        except:
            return val if val else "-"

    def _write_row(self, ws, row_idx, values):
        """배열의 값들을 row_idx 행의 1번 컬럼부터 순서대로 적음. None은 건너뜀"""
        for col_idx, val in enumerate(values):
            if val is not None:
                cell = ws.cell(row=row_idx, column=col_idx+1)
                cell.value = val
                cell.font = self.font_normal
                cell.alignment = self.align_center
                cell.border = self.border_thin
