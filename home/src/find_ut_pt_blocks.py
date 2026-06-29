import os
import openpyxl

desktop = r'C:\Users\-\OneDrive\바탕 화면'
files = [f for f in os.listdir(desktop) if '산출내역서' in f and f.endswith('.xlsx') and not f.startswith('~$')]

if files:
    filepath = os.path.join(desktop, files[0])
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            if 'UT' in sheet_name or 'PT' in sheet_name:
                ws = wb[sheet_name]
                print(f"--- {sheet_name} Headers/Blocks ---")
                for r in range(1, min(100, ws.max_row + 1)):
                    row_data = []
                    for c in range(1, min(20, ws.max_column + 1)):
                        val = ws.cell(row=r, column=c).value
                        if val is not None:
                            if isinstance(val, str) and ('주간' in val or '야간' in val or '휴일' in val or 'UT' in val or 'PT' in val or '가평' in val):
                                row_data.append(str(val).replace('\n', ' '))
                    if row_data:
                        print(f"Row {r}: {row_data}")
    except Exception as e:
        print("Error:", e)
