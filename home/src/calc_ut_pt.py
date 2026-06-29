import os
import openpyxl

desktop = r'C:\Users\-\OneDrive\바탕 화면'
files = [f for f in os.listdir(desktop) if '산출내역서' in f and f.endswith('.xlsx') and not f.startswith('~$')]

def get_sums():
    if not files: return
    filepath = os.path.join(desktop, files[0])
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    for sheet_name in wb.sheetnames:
        if 'UT' in sheet_name or 'PT' in sheet_name:
            ws = wb[sheet_name]
            
            sums = {
                "수송배관(주배관)": {"일반": 0.0, "야간": 0.0, "휴일": 0.0},
                "플랜트(관리소)": {"일반": 0.0, "야간": 0.0, "휴일": 0.0}
            }
            
            current_shift = "일반"
            for r in range(1, ws.max_row + 1):
                row_str = " ".join([str(ws.cell(row=r, column=c).value) for c in range(1, 20) if ws.cell(row=r, column=c).value is not None])
                
                if '주간' in row_str: current_shift = "일반"
                elif '야간' in row_str: current_shift = "야간"
                elif '휴일' in row_str: current_shift = "휴일"
                
                # Check column 5 (1-based is 5) for "주배관" or "관리소"
                val_col5 = str(ws.cell(row=r, column=5).value) if ws.cell(row=r, column=5).value else ""
                
                cat = None
                if '주배관' in val_col5:
                    cat = "수송배관(주배관)"
                elif '관리소' in val_col5:
                    cat = "플랜트(관리소)"
                    
                if cat:
                    # Look for the last numeric value in columns 15 to 20
                    # Usually it's in column 18 for 가산~가평 total
                    val_col18 = ws.cell(row=r, column=18).value
                    val_col16 = ws.cell(row=r, column=16).value
                    
                    # Some rows might not have col18. The total is usually col 18, or if not, col 16 (without correction factor)
                    # Let's take col 18 if it's a number
                    try:
                        amt = float(val_col18)
                        sums[cat][current_shift] += amt
                    except (ValueError, TypeError):
                        pass
            
            print(f"\n[{sheet_name}] Calculated Sums from Excel:")
            for k1, v1 in sums.items():
                for k2, v2 in v1.items():
                    print(f"  {k1} - {k2}: {round(v2, 2)}")

get_sums()
