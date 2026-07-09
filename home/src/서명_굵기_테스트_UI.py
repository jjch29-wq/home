import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageFilter
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import os
import threading

# 전역 변수로 선택된 파일 경로 저장
selected_file_path = ""

def select_file():
    global selected_file_path
    file_path = filedialog.askopenfilename(
        title="도장/서명 이미지 선택",
        filetypes=(("이미지 파일", "*.png *.jpg *.jpeg *.bmp"), ("모든 파일", "*.*"))
    )
    if file_path:
        selected_file_path = file_path
        lbl_file.config(text=f"선택됨: {os.path.basename(file_path)}")

def generate_excel():
    global selected_file_path
    if not selected_file_path:
        messagebox.showwarning("파일 선택", "먼저 굵기를 테스트할 그림(서명) 파일을 선택해 주세요!")
        return

    # 저장할 엑셀 파일 경로 선택창 띄우기
    save_path = filedialog.asksaveasfilename(
        title="엑셀 파일 저장 위치 선택",
        defaultextension=".xlsx",
        initialfile="서명_굵기_비교.xlsx",
        filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")]
    )
    
    if not save_path: # 취소를 누른 경우 그냥 종료
        return

    btn_select.config(state='disabled')
    btn_generate.config(state='disabled')
    lbl_status.config(text="엑셀 파일 생성 중... 잠시만 기다려주세요.", foreground="blue")
    root.update()

    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        signs_dir = os.path.join(script_dir, 'signs')
        output_excel = save_path

        if not os.path.exists(signs_dir):
            os.makedirs(signs_dir)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '서명 굵기 비교'

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20

        headers = ['이름', '1단계 (얇음/원본)', '2단계 (조금 굵게)', '3단계 (중간 굵게)', '4단계 (아주 굵게)']
        for col, text in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=text)
            
        filename = os.path.basename(selected_file_path)

        def make_thick_img(level):
            img = Image.open(selected_file_path).convert('RGBA')
            
            if level == 1:
                pass 
            else:
                data = img.getdata()
                new_data = []
                for item in data:
                    if item[3] > 20: 
                        new_alpha = min(255, int(item[3] * 5)) # 알파값을 진하게 고정
                        new_data.append((0, 0, 0, new_alpha))
                    else:
                        new_data.append((255, 255, 255, 0))
                img.putdata(new_data)
                
                r, g, b, a = img.split()
                # 원본 크기 상태에서 필터를 강하게 주어 굵기를 확실하게 차이 냄
                if level == 2:
                    a = a.filter(ImageFilter.MaxFilter(5))
                elif level == 3:
                    a = a.filter(ImageFilter.MaxFilter(9))
                elif level >= 4:
                    a = a.filter(ImageFilter.MaxFilter(15))
                    
                img = Image.merge('RGBA', (r, g, b, a))
                
            img = img.resize((40, 40), Image.Resampling.LANCZOS)
            temp_path = os.path.join(signs_dir, f'temp_level{level}_custom.png')
            img.save(temp_path, 'PNG')
            return temp_path

        current_row = 2
        ws.row_dimensions[current_row].height = 45
        ws.cell(row=current_row, column=1, value=os.path.splitext(filename)[0])
        
        for level in range(1, 5):
            temp_img_path = make_thick_img(level)
            if temp_img_path:
                xl_img = ExcelImage(temp_img_path)
                col_letter = chr(ord('A') + level)
                ws.add_image(xl_img, f'{col_letter}{current_row}')

        wb.save(output_excel)
        lbl_status.config(text=f"완료! 저장 위치: {output_excel}", foreground="green")
        messagebox.showinfo("생성 완료", f"선택하신 그림의 4단계 굵기 비교 엑셀 파일이 성공적으로 생성되었습니다!\n\n저장 위치:\n{output_excel}")
        
    except Exception as e:
        lbl_status.config(text="오류가 발생했습니다.", foreground="red")
        messagebox.showerror("오류", str(e))
    finally:
        btn_select.config(state='normal')
        btn_generate.config(state='normal')

def on_click_generate():
    threading.Thread(target=generate_excel, daemon=True).start()

# --- UI Setup ---
root = tk.Tk()
root.title("개별 서명 굵기 조절 테스트기")
root.geometry("450x250")
root.resizable(False, False)

style = ttk.Style(root)
style.theme_use('clam')

frame = ttk.Frame(root, padding=20)
frame.pack(fill='both', expand=True)

lbl_title = ttk.Label(frame, text="개별 서명 굵기 엑셀 자동 생성", font=('맑은 고딕', 14, 'bold'))
lbl_title.pack(pady=(0, 10))

lbl_desc = ttk.Label(frame, text="내 컴퓨터에 있는 서명/도장 그림 파일을 선택하면\n그 그림 하나만 4단계 굵기로 변환하여 엑셀로 뽑아줍니다.", justify="center", font=('맑은 고딕', 10))
lbl_desc.pack(pady=(0, 10))

# 파일 선택 버튼 및 라벨
btn_select = ttk.Button(frame, text="📂 그림 파일 찾아보기...", command=select_file, width=25)
btn_select.pack(pady=5)

lbl_file = ttk.Label(frame, text="선택된 파일: 없음", font=('맑은 고딕', 9), foreground="gray")
lbl_file.pack(pady=(0, 10))

btn_generate = ttk.Button(frame, text="엑셀 파일 만들기", command=on_click_generate, width=20)
btn_generate.pack(pady=5)

lbl_status = ttk.Label(frame, text="대기 중...", font=('맑은 고딕', 9))
lbl_status.pack(pady=(5, 0))

root.mainloop()
