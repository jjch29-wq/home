import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.worksheet.pagebreak import Break
import os
from datetime import datetime
import json
from tkcalendar import DateEntry

CONFIG_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "위험성평가_설정.json")

def load_config():
    if os.path.exists(CONFIG_FILE_PATH):
        try:
            with open(CONFIG_FILE_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}

def save_config(config):
    os.makedirs(os.path.dirname(CONFIG_FILE_PATH), exist_ok=True)
    try:
        with open(CONFIG_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=4)
    except:
        pass

# Styles
bold_font = Font(name='맑은 고딕', size=13, bold=True)
title_font = Font(name='맑은 고딕', size=20, bold=True)
data_font = Font(name='맑은 고딕', size=14) # 커진 셀 높이에 어울리도록 기본 글씨 크기 확대
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
center_nowrap_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center", wrap_text=False)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

def set_border(ws, min_col, min_row, max_col, max_row):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border

def get_risk_category(text):
    if any(w in text for w in ["약품", "누출", "화학", "가스", "질식"]): return "화학(물질)적 요인"
    if any(w in text for w in ["방사능", "방사선", "피폭", "소음", "진동", "조도", "폭염"]): return "물리적 요인"
    if any(w in text for w in ["자세", "근골격계", "중량물", "무리한"]): return "작업특성 요인"
    if any(w in text for w in ["우천", "날씨", "법면", "강풍", "태풍", "야외"]): return "작업환경 요인"
    if any(w in text for w in ["감염", "세균", "바이러스", "질병"]): return "생물학적 요인"
    return "기계적(설비)적 요인"

def extract_disaster_type(text):
    if "추락" in text: return "추락"
    if "낙하" in text: return "낙하"
    if "충돌" in text or "부딪" in text: return "충돌"
    if "넘어짐" in text or "미끄러" in text or "전도" in text: return "전도"
    if "근골격" in text: return "근골격계\n질환"
    if "누출" in text or "화상" in text: return "화학물질\n노출"
    if "방사능" in text or "노출" in text or "피폭" in text: return "방사선\n피폭"
    if "붕괴" in text: return "붕괴"
    return "기타"

def create_excel(process_name, output_filename, data, params):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "위험성 평가서"
    
    # 인쇄 방향(가로) 및 자동 맞춤 설정
    ws.page_setup.orientation = 'landscape'
    ws.sheet_view.zoomScale = 75 # 화면 배율
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    # 인쇄 시 페이지 정가운데 맞춤 (좌우, 상하 여백 동일하게)
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    
    ws.print_title_rows = '7:8' # 7~8행 인쇄 제목으로 반복

    # Columns width setup (25 columns: A to Y)
    # 총 너비를 기존보다 아주 살짝 더 넓게(약 261) 설정하여 엑셀 자동 배율이 안전하게 축소되도록 유도 (16행 밀림 방지)
    widths = [12.5, 12, 51, 9, 5, 5, 24] # A:G (C 감소, G 증가)
    widths += [7]                     # H
    widths += [8.2] * 7               # I~O
    widths += [7]                     # P
    
    # 우측 '재평가' 블록(Q~Y)
    widths += [7]                     # Q
    widths += [8.2, 8.2, 12, 8.2, 8.2, 8.2, 8.2] # R~X (T열은 담당부서 글자 길이 때문에 12로 확대)
    widths += [7]                     # Y (원래 비율에 가깝게 복구)
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # --- Header Section (Rows 1 to 6) ---
    ws.merge_cells('A1:Y1')
    eval_type = params['eval_type']
    if eval_type == "최초평가":
        ws['A1'] = "위험성평가표 [ ■최초, □정기, □수시 ]"
    elif eval_type == "정기평가":
        ws['A1'] = "위험성평가표 [ □최초, ■정기, □수시 ]"
    else:
        ws['A1'] = "위험성평가표 [ □최초, □정기, ■수시 ]"
    ws['A1'].font = Font(name='맑은 고딕', size=18, bold=True)
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 40

    # Left Block
    ws.merge_cells('A2:B3'); ws['A2'] = "작 업 명"
    ws.merge_cells('C2:G3'); ws['C2'] = f"{params['site_name']}\n{process_name} 기술용역"
    ws.merge_cells('A4:B4'); ws['A4'] = "회 사 명"
    ws.merge_cells('C4:G4'); ws['C4'] = params['company_name']
    ws.merge_cells('A5:B5'); ws['A5'] = "작업공종"
    ws.merge_cells('C5:G5'); ws['C5'] = process_name
    ws.merge_cells('A6:B6'); ws['A6'] = "작업기간"
    ws.merge_cells('C6:G6'); ws['C6'] = f"{params['start_date']} ~ {params['end_date']}"

    # Middle Block
    ws.merge_cells('H2:I2'); ws['H2'] = "평가기간"
    ws.merge_cells('J2:P2'); ws['J2'] = params['write_date']
    ws.merge_cells('H3:H6'); ws['H3'] = "작업전"
    
    ws.merge_cells('I3:I4'); ws['I3'] = "수급인"
    ws.merge_cells('J3:K3'); ws['J3'] = "근 로 자"
    ws.merge_cells('L3:M3'); ws['L3'] = "작 성 자"
    ws.merge_cells('N3:P3'); ws['N3'] = "승 인 자"
    ws.merge_cells('J4:K4'); ws['J4'] = params.get('worker_name', '(서명)')
    ws.merge_cells('L4:M4'); ws['L4'] = params.get('writer_name', '(서명)')
    ws.merge_cells('N4:P4'); ws['N4'] = params.get('approver_name', '(서명)')
    
    ws.merge_cells('I5:I6'); ws['I5'] = "도급인"
    ws.merge_cells('J5:K5'); ws['J5'] = "검토자(감독)"
    ws.merge_cells('L5:M5'); ws['L5'] = "검토자(안전)"
    ws.merge_cells('N5:P5'); ws['N5'] = "승 인 자"
    ws.merge_cells('J6:K6'); ws['J6'] = "(서명)"
    ws.merge_cells('L6:M6'); ws['L6'] = "(서명)"
    ws.merge_cells('N6:P6'); ws['N6'] = "(서명)"

    # Right Block
    ws.merge_cells('Q2:R2'); ws['Q2'] = "재평가일"
    ws.merge_cells('S2:Y2'); ws['S2'] = ""
    ws.merge_cells('Q3:Q6'); ws['Q3'] = "재평가"
    
    ws.merge_cells('R3:R4'); ws['R3'] = "수급인"
    ws.merge_cells('S3:T3'); ws['S3'] = "근 로 자"
    ws.merge_cells('U3:V3'); ws['U3'] = "작 성 자"
    ws.merge_cells('W3:Y3'); ws['W3'] = "승 인 자"
    ws.merge_cells('S4:T4'); ws['S4'] = "(서명)"
    ws.merge_cells('U4:V4'); ws['U4'] = "(서명)"
    ws.merge_cells('W4:Y4'); ws['W4'] = "(서명)"
    
    ws.merge_cells('R5:R6'); ws['R5'] = "도급인"
    ws.merge_cells('S5:T5'); ws['S5'] = "검토자(감독)"
    ws.merge_cells('U5:V5'); ws['U5'] = "검토자(안전)"
    ws.merge_cells('W5:Y5'); ws['W5'] = "승 인 자"
    ws.merge_cells('S6:T6'); ws['S6'] = "(서명)"
    ws.merge_cells('U6:V6'); ws['U6'] = "(서명)"
    ws.merge_cells('W6:Y6'); ws['W6'] = "(서명)"

    for r in range(2, 7):
        if r in [4, 6]:
            ws.row_dimensions[r].height = 40
        else:
            ws.row_dimensions[r].height = 30
        for c in range(1, 26):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center_align
            cell.font = bold_font
            
    # 특정 셀 줄바꿈 방지 (작업전, 재평가 및 수급인/도급인 통일)
    ws['H3'].alignment = center_nowrap_align
    ws['Q3'].alignment = center_nowrap_align
    ws['I3'].alignment = center_nowrap_align
    ws['I5'].alignment = center_nowrap_align
    ws['R3'].alignment = center_nowrap_align
    ws['R5'].alignment = center_nowrap_align
    
    # 서명(이름) 칸 세부 정렬 설정 및 도장 이미지 자동 삽입
    # 승인자 칸(N~P열)처럼 셀이 넓은 경우 양쪽 끝으로 밀어버리면 간격이 너무 벌어지므로,
    # 고정된 간격(스페이스 6개)을 주고 통째로 가운데 정렬하여 하나의 도장처럼 보이게 합니다.
    for cell_ref in ['J4', 'L4', 'N4', 'J6', 'L6', 'N6', 'S4', 'U4', 'W4', 'S6', 'U6', 'W6']:
        cell_val = ws[cell_ref].value
        if cell_val:
            name = str(cell_val).replace(" (서명)", "").strip()
            
            # 도장 이미지 확인
            stamp_path = None
            if name in ["유상훈", "주진철", "강신태"]:
                ext = "_padded.png"
                temp_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "signs", f"{name}{ext}")
                if os.path.exists(temp_path):
                    stamp_path = temp_path

            if not name or name == "(서명)":
                ws[cell_ref].value = "(서명)"
                ws[cell_ref].alignment = right_align
            else:
                if stamp_path:
                    # 도장이 삽입될 때 이름과 도장 간의 간격을 정밀하게 조절합니다.
                    # 좁은 셀(J, L 등)은 가운데 정렬 시 텍스트가 밀려 도장과 겹치므로 공백을 4칸 주어 적당한 간격을 만듭니다.
                    # 넓은 셀(N, W 등)은 도장이 P열에 고정되므로, 공백을 2칸으로 줄여 이름이 도장에 더 가까워지도록 합니다.
                    if cell_ref.startswith('N') or cell_ref.startswith('W'):
                        ws[cell_ref].value = f"{name}  "
                    else:
                        ws[cell_ref].value = f"{name}      "
                    ws[cell_ref].alignment = center_nowrap_align
                    
                    # 도장 이미지 삽입
                    img = Image(stamp_path)
                    img.width = 85 if name == "강신태" else 55
                    img.height = 40
                    
                    # 이미지 삽입 위치 (강신태는 너무 멀지 않게 P대신 O열, Y대신 X열을 기준으로 당김)
                    if cell_ref.startswith('J'): target_col = 'K'
                    elif cell_ref.startswith('L'): target_col = 'M'
                    elif cell_ref.startswith('N'): target_col = 'O'
                    elif cell_ref.startswith('S'): target_col = 'T'
                    elif cell_ref.startswith('U'): target_col = 'V'
                    elif cell_ref.startswith('W'): target_col = 'X'
                    else: target_col = chr(ord(cell_ref[0]) + 1)
                    
                    target_ref = f"{target_col}{cell_ref[1:]}"
                    ws.add_image(img, target_ref)
                else:
                    # 도장이 없는 경우 기존처럼 텍스트로 (서명) 삽입
                    if cell_ref.startswith('N') or cell_ref.startswith('W'):
                        ws[cell_ref].value = f"{name}      (서명)"
                    else:
                        ws[cell_ref].value = f"{name}  (서명)"
                    ws[cell_ref].alignment = center_nowrap_align

    set_border(ws, 1, 2, 25, 6)

    # --- Table Headers (Rows 7 & 8) ---
    ws.merge_cells('A7:A8'); ws['A7'] = "세부작업"
    ws.merge_cells('B7:B8'); ws['B7'] = "위험분류"
    
    ws.merge_cells('C7:F7'); ws['C7'] = "유해 위험 요인 파악"
    ws.merge_cells('C8:F8'); ws['C8'] = "위험 발생 상황 및 결과"
    
    ws.merge_cells('G7:I8'); ws['G7'] = "현재의 안전 보건 조치"
    
    ws.merge_cells('J7:L7'); ws['J7'] = "위험성 결정"
    ws['J8'] = "가능성\n(빈도)"
    ws['K8'] = "중대성\n(강도)"
    ws['L8'] = "위험성\n(등급)"
    
    ws.merge_cells('M7:Q8'); ws['M7'] = "위험성 감소 대책"
    
    ws.merge_cells('R7:U7'); ws['R7'] = "재평가"
    ws.merge_cells('V7:Y7'); ws['V7'] = "개선 후 위험성 (재평가)"
    
    ws.merge_cells('R8:S8'); ws['R8'] = "개선예정일"
    ws['T8'] = "담당자\n(담당부서)"
    ws['U8'] = "완료일"
    ws['V8'] = "가능성\n(빈도)"
    ws['W8'] = "중대성\n(강도)"
    ws.merge_cells('X8:Y8'); ws['X8'] = "위험성\n(등급)"

    for r in range(7, 9):
        ws.row_dimensions[r].height = 35
        for c in range(1, 26):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center_align
            cell.font = bold_font
            cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            
    # 특정 셀(현재의 안전 보건 조치) 줄바꿈 방지
    ws['G7'].alignment = center_nowrap_align

    set_border(ws, 1, 7, 25, 8)

    # --- Data Definition ---
    start_row = 9
    current_row = 9

    a_groups = []
    current_a = None
    current_a_count = 0
    for item in data:
        cat = item[0]
        if current_a == cat:
            current_a_count += 1
        else:
            if current_a is not None:
                a_groups.append((current_a, current_a_count))
            current_a = cat
            current_a_count = 1
    if current_a is not None:
        a_groups.append((current_a, current_a_count))
        
    s_date_short = params['start_date'].replace("년 ", ".").replace("월 ", ".").replace("일", "")
    e_date_short = params['end_date'].replace("년 ", ".").replace("월 ", ".").replace("일", "")

    for item in data:
        freq = item[3]
        sev = item[4]
        
        # 작업이 "허용가능"하도록 점수 최대 3점(VI등급) 이하로 제한 (고객 요청: 1~3점 이상이면 안됨)
        try:
            f_val = int(freq)
            s_val = int(sev)
            while f_val * s_val > 3:
                if f_val >= s_val:
                    f_val -= 1
                else:
                    s_val -= 1
            freq = f_val
            sev = s_val
            
            # 첨부해주신 실제 위험성 등급표 기준에 맞춘 등급 계산
            score = freq * sev
            if score >= 16:
                grade = "I"
            elif score >= 15:
                grade = "II"
            elif score >= 9:
                grade = "III"
            elif score >= 8:
                grade = "IV"
            elif score >= 4:
                grade = "V"
            else:
                grade = "VI"
                
        except (ValueError, TypeError):
            grade = item[5] if len(item) > 5 else ""
            
        ws.cell(row=current_row, column=1, value=item[0]).alignment = center_align
        ws.cell(row=current_row, column=2, value=get_risk_category(item[1])).alignment = center_align
        
        ws.merge_cells(f'C{current_row}:F{current_row}')
        ws.cell(row=current_row, column=3, value=item[1]).alignment = left_align
        
        ws.merge_cells(f'G{current_row}:I{current_row}')
        ws.cell(row=current_row, column=7, value=item[2] if len(item) > 2 else "").alignment = left_align
        
        ws.cell(row=current_row, column=10, value=freq).alignment = center_align
        ws.cell(row=current_row, column=11, value=sev).alignment = center_align
        ws.cell(row=current_row, column=12, value=grade).alignment = center_align
        
        # 위험성 점수가 1~3점(VI등급) 이하인 경우 위험성 감소 대책을 숨김(빈칸 처리)
        try:
            score_val = int(freq) * int(sev)
        except:
            score_val = 0
            
        if score_val <= 3:
            countermeasure = ""
        else:
            countermeasure = item[6] if len(item) > 6 else ""
            
        ws.merge_cells(f'M{current_row}:Q{current_row}')
        ws.cell(row=current_row, column=13, value=countermeasure).alignment = left_align
        
        ws.merge_cells(f'R{current_row}:S{current_row}')
        ws.cell(row=current_row, column=18, value="").alignment = center_align
        ws.cell(row=current_row, column=20, value="").alignment = center_align
        ws.cell(row=current_row, column=21, value="").alignment = center_align
        ws.cell(row=current_row, column=22, value="").alignment = center_align
        ws.cell(row=current_row, column=23, value="").alignment = center_align
        
        ws.merge_cells(f'X{current_row}:Y{current_row}')
        ws.cell(row=current_row, column=24, value="").alignment = center_align
        
        ws.row_dimensions[current_row].height = 93
        current_row += 1

    # 마지막 페이지 빈 공간을 빈 행으로 채워서 높이를 꽉 채움 (위쪽 정렬 효과)
    last_data_row = current_row - 1
    if last_data_row <= 16:
        target_row = 16
        total_pages = 1
    else:
        page_idx = (last_data_row - 17) // 10
        target_row = 16 + (page_idx + 1) * 10
        total_pages = page_idx + 2
        
    # 인쇄 및 PDF 변환 시 엑셀이 멋대로 여백을 늘려 빈 페이지를 만들지 못하도록 
    # 전체 페이지 수(세로 높이)를 정확한 숫자로 못박아 둡니다.
    ws.page_setup.fitToHeight = 0

    # 인쇄물에서 비율 차이로 인해 17행이 1페이지로 빨려들어가는 현상을 방지하기 위해
    # 각 페이지가 끝나는 행에 '강제 페이지 나누기(Page Break)'를 삽입합니다.
    for p_idx in range(total_pages - 1):
        break_row = 16 + (p_idx * 10)
        ws.row_breaks.append(Break(id=break_row))

    while current_row <= target_row:
        ws.row_dimensions[current_row].height = 93
        ws.merge_cells(f'C{current_row}:F{current_row}')
        ws.merge_cells(f'G{current_row}:I{current_row}')
        ws.merge_cells(f'M{current_row}:Q{current_row}')
        ws.merge_cells(f'R{current_row}:S{current_row}')
        ws.merge_cells(f'X{current_row}:Y{current_row}')
        current_row += 1

    # 세부작업, 위험분류 세로 병합 (데이터 값 기준으로 병합 그룹핑 변경)
    r = start_row
    # item[0] (세부작업)과 item[1] 기반 위험분류가 같으면 같이 병합되도록 묶기
    b_groups = []
    current_b = None
    current_b_count = 0
    for item in data:
        cat_a = item[0]
        cat_b = get_risk_category(item[1])
        pair = (cat_a, cat_b)
        
        if current_b == pair:
            current_b_count += 1
        else:
            if current_b is not None:
                b_groups.append((current_b, current_b_count))
            current_b = pair
            current_b_count = 1
    if current_b is not None:
        b_groups.append((current_b, current_b_count))

    # A열(세부작업) 병합
    a_r = start_row
    for cat, count in a_groups:
        if count > 1:
            current_start = a_r
            current_end = a_r + count - 1
            first_page_break = 16 # 16행까지 1페이지
            while current_start <= current_end:
                if current_start <= first_page_break:
                    page_end = first_page_break
                else:
                    page_idx = (current_start - (first_page_break + 1)) // 10
                    page_end = first_page_break + (page_idx + 1) * 10
                merge_end = min(current_end, page_end)
                
                # 인쇄 페이지가 넘어가서 병합이 끊어지는 경우, 다음 페이지 첫 셀에 내용 다시 입력
                if current_start > a_r:
                    ws.cell(row=current_start, column=1, value=cat).alignment = center_align
                
                if current_start < merge_end:
                    ws.merge_cells(f'A{current_start}:A{merge_end}')
                current_start = merge_end + 1
        a_r += count
        
    # B열(위험분류) 병합
    b_r = start_row
    for pair, count in b_groups:
        cat_b = pair[1]
        if count > 1:
            current_start = b_r
            current_end = b_r + count - 1
            first_page_break = 16
            while current_start <= current_end:
                if current_start <= first_page_break:
                    page_end = first_page_break
                else:
                    page_idx = (current_start - (first_page_break + 1)) // 10
                    page_end = first_page_break + (page_idx + 1) * 10
                merge_end = min(current_end, page_end)
                
                # 인쇄 페이지가 넘어가서 병합이 끊어지는 경우, 다음 페이지 첫 셀에 내용 다시 입력
                if current_start > b_r:
                    ws.cell(row=current_start, column=2, value=cat_b).alignment = center_align
                
                if current_start < merge_end:
                    ws.merge_cells(f'B{current_start}:B{merge_end}')
                current_start = merge_end + 1
        b_r += count

    set_border(ws, 1, 7, 25, current_row - 1)
    
    # 커진 셀 높이(105)에 맞춰 9행 이하 데이터 셀들의 글자 크기를 최대 16pt로 설정
    for row in ws.iter_rows(min_row=9, max_row=current_row - 1, min_col=1, max_col=25):
        for cell in row:
            current_font_size = 16 # 기본 글자 크기를 16pt로 대폭 확대
            if cell.value and isinstance(cell.value, str):
                lines = cell.value.split('\n')
                
                # 열 너비에 따라 한 줄에 들어가는 글자 수가 다르므로 각각 다르게 계산
                col_idx = cell.column
                if col_idx in [3, 4, 5, 6]: # C~F열 (유해 위험 요인 파악 - 너비 85)
                    estimated_lines = sum((len(line) // 40) + 1 for line in lines)
                elif col_idx in [13, 14, 15, 16, 17]: # M~Q열 (위험성 감소 대책 - 너비 약 38)
                    estimated_lines = sum((len(line) // 22) + 1 for line in lines)
                else:
                    estimated_lines = len(lines)
                
                # 105 높이에서 글자가 잘리지 않도록 줄 수에 따라 크기 대폭 축소 (특히 21행 M~Q열 등)
                if estimated_lines >= 7:
                    current_font_size = 11
                elif estimated_lines >= 6:
                    current_font_size = 12 # 21행 등 내용이 많은 곳은 12pt로 고정
                elif estimated_lines >= 5:
                    current_font_size = 14
                    
            cell.font = Font(name='맑은 고딕', size=current_font_size)

    # 인쇄 영역을 정확하게 지정하여 불필요한 빈 페이지(4페이지 등)가 인쇄/PDF 변환되지 않도록 방지
    ws.print_area = f'A1:Y{current_row - 1}'

    try:
        wb.save(output_filename)
        return True, f"[{process_name}] 저장 완료: {os.path.basename(output_filename)}"
    except Exception as e:
        return False, f"Error saving {output_filename}: {e}"

# Data Definitions
data_rt = [
    ["이동", "이동 중 지면 단차, 돌출물, 자재 등에 걸려 넘어짐 (전도) 위험", "개인 랜턴으로 통로 조도 확보 / 바닥 장애물 사전 제거", 2, 2, "IV", "1. 가설계단 및 이동통로 내 자재 정리정돈 철저\n2. 조도(밝기) 확보 및 보행 중 스마트폰 사용 금지 전파"],
    ["이동", "이동용 사다리를 이용한 고소 이동 중 사다리 전도 및 추락", "안전모 착용 / 2인 1조 작업 / 아웃트리거 사용", 3, 3, "V", "1. 사다리 상단 고정 및 아웃트리거 전개 확인\n2. 사다리 이용 시 반드시 2인 1조 작업"],
    ["이동", "현장 내 중장비(굴착기 등) 주변 이동 중 장비와 충돌 또는 끼임", "중장비 선회 반경 내 절대 접근 금지 및 우회 통행 철저", 2, 3, "V", "1. 중장비 가동 중 작업 반경 내 진입 절대 금지 및 우회 통로 이용\n2. 부득이하게 장비 주변 이동 시 운전원과 Eye Contact 및 수신호 교환"],
    ["사전준비", "필름부착 작업 시 배관상부에서 미끄러져 넘어짐 위험", "안전대 체결 불가 시 작업화 이물질 제거 및 2인 1조 밀착 통제", 3, 2, "V", "1. 배관 탑승 시 하부에서 1인이 지속 모니터링 및 밀착 지지\n2. 우천/결빙 시 배관 상부 탑승 작업 전면 중지"],
    ["사전준비", "지하(트렌치) 작업 중 상부 낙하물에 의한 사고위험", "하부 진입 전 작업자가 상단부 낙하위험물 직접 점검 및 제거", 2, 3, "V", "1. 굴착 상단부에 자재 및 공구 적재 금지\n2. 하부 작업 시 상부 타 공정 동시 작업 금지"],
    ["사전준비", "필름부착작업 시 사다리 작업 중 추락/전도 위험", "최상단 탑승 금지 자체 확인 / 사다리 2인 1조 작업(1인 지지)", 3, 2, "V", "1. A형 사다리 최상단 발판 탑승 절대 금지\n2. 1인 하단부 지지 (2인 1조 작업)"],
    ["사전준비", "차폐체 설치/해체 작업 시 중량물 낙하에 의한 사고 위험", "납 차폐체 이동 시 전용 대차 사용 / 2인 이상 협력하여 운반", 2, 3, "V", "1. 100kg 이상 인양 시 2줄 걸이 이상 결속\n2. 인양 중인 중량물 하부 통행 전면 금지"],
    ["사전준비", "인소스로봇(크롤러) 투입/반출 작업 시 낙하 위험", "인양 전 슬링벨트 상태 육안 점검 및 결속부 2중 자체 확인", 2, 3, "V", "1. 투입 전용 인양 장비 사용 및 결속 상태 2중 확인\n2. 배터리 및 조작계 이상 시 배관 내부 인력 투입 금지"],
    ["사전준비", "손상된 슬링벨트 사용 및 줄걸이 부적합으로 낙하 사고 위험", "작업 전 슬링벨트 자체 점검(손상품 즉시 폐기) 및 2줄 걸이 결속 확인", 2, 3, "V", "1. 작업 전 슬링벨트 손상 규격 확인 및 파손품 폐기\n2. 관리감독자 지휘 하에 안전하게 인양 진행"],
    ["사전준비", "협소공간 필름부착 시 부자연스런 자세에 의한 근골격계질환", "작업 전후 스트레칭 및 중량물 취급 시 올바른 자세 유지", 2, 2, "IV", "1. 작업 전·후 스트레칭 실시 및 중량물 인양 올바른 자세 교육\n2. 장시간 작업 시 교대 및 휴식 시간 부여"],
    ["사전준비", "우천 등으로 인한 굴착법면 상태 약화로 붕괴 위험", "진입 전 굴착 법면 균열 및 용수(물비침) 육안 점검", 2, 3, "V", "1. 작업 전 굴착 법면 균열, 용수(물비침) 점검\n2. 호우 경보 발효 시 굴착 지점 내부 진입 금지"],
    ["비파괴검사", "필름 현상 작업 시 약품 취급 오류/누출에 의한 흡입 및 화상 위험", "작업 전 방독마스크, 내화학 장갑 자체 착용 상태 교차 점검", 2, 3, "V", "1. 현상액/정착액 MSDS(물질안전보건자료) 비치 및 교육\n2. 작업 시 방독마스크, 내화학 장갑 착용 철저"],
    ["비파괴검사", "검사안전구역 미설정 및 통제 부실 시 방사능 노출 피폭 위험", "출입금지 간판 설치", 3, 3, "V", "1. 콜리메이터(차폐장비) 필수 장착으로 선량 최소화\n2. 방사선 10μSv/hr 이하 통제구역 설정 및 전담 감시자 배치\n3. 작업자 알람메타/개인선량계 항시 패용 (경보 시 대피)"],
    ["작업 후 정리", "필름 및 차폐체 해체 작업 시 배관상부 미끄러짐 및 추락/낙하", "해체 자재는 하부 조원에게 직접 수작업 인계 / 자재 투척 절대 금지", 3, 2, "V", "1. 배관 하강 완료 전까지 안전모 턱끈 결속 유지\n2. 해체된 자재는 2인 1조로 안전하게 수작업 인계 및 던지기 엄금"],
    ["작업 후 정리", "웰딩하우스 하부 및 현장 정리 중 두부 충돌 및 부딪힘 위험", "진입 전 헤드랜턴으로 상부 장애물 사전 확인 및 안전모 턱끈 결속 철저", 2, 2, "IV", "1. 웰딩하우스 진입 전 헤드랜턴으로 상부 돌출물 육안 확인\n2. 철수 시에도 안전모 미착용 절대 금지"],
    ["비파괴검사", "가이드튜브 꺾임, 장비 결함 등으로 동위원소 선원 미회수 시 고선량 피폭 위험", "장비 육안 점검 / 현장에 비상용 롱 텅(Handling tong) 사전 비치", 1, 3, "III", "1. 현장에 비상용 롱 텅(Handling tong) 및 납 차폐체 상시 비치\n2. 비상사태 발생 시 방사선안전관리자 즉시 보고 및 비상대응 매뉴얼 숙지"],
    ["비파괴검사", "야간 검사 시 조명 부족 및 시야 미확보로 인한 전도, 추락, 장비 충돌 위험", "헤드랜턴 및 고휘도 반사조끼 착용", 2, 2, "IV", "1. 검사 구역 및 주요 이동 통로에 30 Lux 이상의 가설 투광기 추가 설치\n2. 작업자 전원 고휘도 반사조끼 착용 및 야광 테이프로 걸림 위험 부위 표시"],
    ["사전준비", "배관 내부 및 깊은 트렌치 내부 출입 시 산소 결핍 또는 유해가스 체류로 인한 질식", "배관 내부 인력 진입 절대 금지(크롤러 장비 대체) / 트렌치 진입 전 복합가스 자체 측정", 2, 3, "V", "1. 배관 내부 진입 절대 금지 및 크롤러 등 장비로 대체\n2. 트렌치 진입 전 복합가스 측정 및 외부 감시인 배치"],
    ["비파괴검사", "야간 조명등 및 검사장비 케이블 피복 손상, 누전, 우천 시 감전 위험", "누전차단기 부착형 릴선 전용 사용 및 손상된 케이블 즉시 폐기", 2, 2, "IV", "1. 현장 내 모든 임시 가설전기는 누전차단기(30mA)를 거쳐서 사용\n2. 피복 손상 케이블 즉각 폐기 및 바닥 띄움 거치 실시"],
    ["이동", "야간 암실 차량 주·정차 및 이동 시 시야 확보 불량으로 근로자 및 타 장비 충돌", "동승자 하차 후 신호봉 자체 직접 차량 유도", 2, 3, "V", "1. 암실차 주·정차 시 점멸등(윙카) 작동 및 작업자 고휘도 반사조끼 착용\n2. 야간 차량 후진 및 좁은 구역 이동 시 동승자(작업자)가 하차하여 신호봉으로 직접 유도"],
    ["작업 후 정리", "차량 이동 중 동위원소 저장용기 전도/낙하로 인한 방사능 유출 및 장비 파손", "차량 내 저장용기 자체 이중 시건장치 확인", 1, 3, "III", "1. 차량 내 저장용기 이중 시건장치 결속 및 와이어로프로 차체에 단단히 고정\n2. 운송 중 과속/급정거 엄금 및 방사선 경고 표지 차량 4면 부착"]
]

data_ut = [
    ["사전준비", "지하 작업 중 낙하물에 의한 사고위험", "하부 진입 전 작업자가 상단부 낙하위험물 직접 확인 및 제거", 2, 3, "V", "1. 굴착 상단부에 자재 및 공구 적재 금지\n2. 하부 작업 시 상부 타 공정 혼재 작업 원칙적 금지"],
    ["사전준비", "소형 장비(탐상기 등) 수작업 운반 중 계단/비계에서 추락 및 전도", "전용 수납가방 활용으로 양손 확보 및 3점 지지 승강 철저", 2, 2, "IV", "1. 장비 전용 수납가방 사용으로 이동 시 양손 확보\n2. 사다리/비계 승강 시 3점 지지 철저"],
    ["사전준비", "배관 상부 등 고소작업 시 소형 장비(탐상기, 겔 등) 하부 낙하", "장비 낙하방지끈 체결 및 하부 작업자 접근 통제", 2, 2, "IV", "1. 배관 탑승 작업 시 탐상기 및 공구 낙하방지끈 체결\n2. 작업 구역 하부 근로자 접근 전면 통제"],
    ["사전준비", "근로자가 안전모 등 개인 보호구 미착용에 의한 충돌", "현장 투입 전 개인보호구 착용 및 안전모 턱끈 자체 체결 확인", 2, 2, "V", "1. 작업구역 진입 전 개인보호구 착용 상태 점검\n2. 안전모 턱끈 결속 철저"],
    ["사전준비", "중장비 주변 이동 중 장비와 충돌 또는 끼임", "운전원과 직접 Eye Contact 및 수신호 교환 후 통행", 2, 2, "V", "1. 장비 가동 중 작업 반경 내 진입 절대 금지 및 우회 통로 이용\n2. 부득이하게 장비 주변 이동 시 운전원과 눈맞춤 및 수신호 확인 필수"],
    ["비파괴검사", "협소공간 검사 시 부자연스런 자세에 의한 근골격계질환 위험", "작업 전 스트레칭", 2, 1, "VI", "1. 작업 전·후 전신 스트레칭 실시\n2. 장시간 구부린 자세 유지 시 주기적인 휴식(10분 이상) 부여"],
    ["비파괴검사", "우천으로 인한 굴착법면상태 악화에 의한 붕괴 위험", "지하 진입 전 법면 균열/용수(물비침) 자체 육안 점검", 2, 3, "V", "1. 작업 전 굴착 법면 균열, 용수(물비침) 현상 일일 점검\n2. 호우 경보 발효 시 굴착 지점 내부 진입 전면 금지"],
    ["작업 후 정리", "검사 완료 후 장비 반출 시 계단/비계에서 추락 및 전도", "장비 전용 가방 수납 후 양손 확보 상태로 안전하게 하강", 2, 2, "IV", "1. 반출 시 장비를 손에 쥐고 이동하는 행위 절대 금지\n2. 전용 가방 수납 후 3점 지지 하강 원칙 준수"],
    ["작업 후 정리", "장비 반출 중 수작업 인계 시 하부로 장비 낙하", "하부 조원에게 직접 수작업 인계 및 장비 투척 절대 금지", 2, 2, "IV", "1. 2인 1조 작업으로 하부 조원에게 안전하게 수작업 인계\n2. 장비를 바닥으로 휙 던지는 행위 절대 금지"],
    ["작업 후 정리", "중장비 주변 이동 및 철수 중 장비와 충돌사고 위험", "운전원과 수신호 확인 및 자체 동승자 하차 차량 유도", 2, 2, "V", "1. 장비 철수 동선 파악 및 작업 반경 밖으로 우회 통행 철저\n2. 장비 주변 비파괴 자체 차량 이동 시 동승자가 하차하여 직접 유도"]
]

data_pt = [
    ["사전준비", "지하 작업 중 낙하물에 의한 사고위험", "하부 진입 전 작업자가 상단부 낙하위험물 직접 확인 및 제거", 2, 3, "V", "1. 굴착 상단부에 자재 및 공구 적재 금지\n2. 하부 작업 시 상부 타 공정 혼재 작업 원칙적 금지"],
    ["사전준비", "소형 자재(스프레이, 세척액 등) 수작업 운반 중 추락/전도", "전용 수납가방 활용으로 양손 확보 및 3점 지지 승강 철저", 2, 2, "IV", "1. 스프레이 등 탐상 자재 전용 포켓/가방에 수납\n2. 사다리/비계 승강 시 3점 지지 철저"],
    ["사전준비", "배관 상부 등 고소작업 시 스프레이 캔 하부 낙하", "작업 구역 하부 접근 통제 및 잔여 자재 안전한 곳에 비치", 2, 2, "IV", "1. 비계나 배관 위 작업 시 스프레이가 굴러떨어지지 않도록 조치\n2. 하부 근로자 통행 전면 금지"],
    ["사전준비", "근로자가 안전모 등 개인 보호구 미착용에 의한 충돌", "현장 투입 전 개인보호구 착용 및 안전모 턱끈 자체 체결 확인", 2, 2, "V", "1. 작업구역 진입 전 개인보호구 착용 상태 점검\n2. 안전모 턱끈 결속 철저"],
    ["사전준비", "중장비 주변 이동 중 장비와 충돌 또는 끼임", "운전원과 직접 Eye Contact 및 수신호 교환 후 통행", 2, 2, "V", "1. 장비 가동 중 작업 반경 내 진입 절대 금지 및 우회 통로 이용\n2. 부득이하게 장비 주변 이동 시 운전원과 눈맞춤 및 수신호 확인 필수"],
    ["비파괴검사", "대상 화학물질에 대한 유해위험성 미인식에 의한 건강장해 발생 위험", "침투탐상제 취급 전 MSDS 유해성 자체 숙지 및 전파", 1, 1, "VI", "1. 침투탐상제(세척액, 침투액, 현상액) MSDS 비치\n2. 작업 전 특별 안전보건 교육 실시 및 취급 주의 전파"],
    ["비파괴검사", "화학물질 취급작업시 보호구 미착용으로 인한 사고 위험", "화학물질 취급 전 내화학 장갑/마스크 자체 수령 및 착용", 1, 1, "VI", "1. 내화학 장갑 및 방독마스크 지급 및 착용 의무화\n2. 보호구 훼손 시 즉시 새 제품으로 교체 지급"],
    ["비파괴검사", "협소공간 검사 시 부자연스런 자세에 의한 근골격계질환 위험", "작업 전 스트레칭", 2, 1, "VI", "1. 작업 전·후 전신 스트레칭 실시\n2. 장시간 구부린 자세 유지 시 주기적인 휴식 부여"],
    ["비파괴검사", "우천으로 인한 굴착법면상태 악화에 의한 붕괴 위험", "지하 진입 전 법면 균열/용수(물비침) 자체 육안 점검", 2, 3, "V", "1. 작업 전 굴착 법면 균열, 용수(물비침) 일일 점검\n2. 호우 경보 시 굴착 지점 내부 진입 전면 금지"],
    ["작업 후 정리", "검사 완료 후 자재 반출 시 계단/비계에서 추락 및 전도", "자재를 전용 가방 수납 후 양손 확보 상태로 안전하게 하강", 2, 2, "IV", "1. 다 쓴 스프레이 캔 등을 손에 쥐고 이동 금지\n2. 빈 캔은 가방에 회수 후 3점 지지 하강 원칙 준수"],
    ["작업 후 정리", "자재 반출 중 수작업 인계 시 하부로 스프레이 캔 낙하", "하부 조원에게 직접 수작업 인계 및 폐캔 투척 절대 금지", 2, 2, "IV", "1. 하부 조원에게 빈 캔을 건네줄 때 안전하게 손으로 인계\n2. 사용이 끝난 빈 캔 바닥 투척 절대 금지"],
    ["작업 후 정리", "중장비 주변 이동 및 철수 중 장비와 충돌사고 위험", "운전원과 수신호 확인 및 자체 동승자 하차 차량 유도", 2, 2, "V", "1. 장비 철수 동선 파악 및 작업 반경 밖으로 우회 통행 철저\n2. 장비 주변 비파괴 자체 차량 이동 시 동승자가 하차하여 직접 유도"]
]

data_container = [
    ["반입 및 설치", "크레인 양중 작업 중 줄걸이(슬링벨트) 파단 및 체결 불량으로 낙하", "작업 전 슬링벨트 자체 점검(손상품 즉시 폐기) 및 4줄 걸이 결속 확인", 2, 3, "V", "1. 중량물 취급 작업계획서 작성 및 지정된 신호수 배치\n2. 4줄 걸이 양중 원칙 준수 및 인양물 하부 출입 통제"],
    ["반입 및 설치", "설치 지반의 평탄성 불량 및 침하로 인한 컨테이너 전도", "작업 전 점검", 2, 3, "V", "1. 설치 전 지반 평탄화 작업 및 단단한 지지대(고임목) 설치\n2. 강풍 대비 와이어로프 결속(타이다운) 조치"],
    ["전기설비", "전원 연결 시 규격 미달 전선 사용 및 접지 불량으로 인한 감전", "전원 연결부 자체 접지 상태 육안 확인 및 누전차단기 테스트", 2, 3, "V", "1. 메인 분전반 내 누전차단기 설치 및 정상 작동 여부 확인\n2. 가설전기 외함 접지(3종) 실시 및 전선 바닥 띄움 거치"],
    ["전기설비", "문어발식 콘센트 사용 및 전열기구 과열로 인한 화재 발생", "소화기 비치", 2, 3, "V", "1. 컨테이너 내부 문어발식 콘센트 사용 금지 및 정격 용량 준수\n2. 내부 소화기(3.3kg 이상) 비치 및 화재경보기 설치"],
    ["유지 및 운영", "컨테이너 출입구 계단 단차 및 결빙으로 인한 미끄러짐(전도)", "미끄럼 방지 테이프 자체 확인 및 결빙 시 제설 조치 철저", 2, 2, "V", "1. 출입 계단 미끄럼 방지 테이프 부착 및 안전 난간대 설치\n2. 우천/결빙 시 모래함 비치 및 제설 작업 철저"],
    ["유지 및 운영", "환기 불량 상태에서 난방기기 사용 중 일산화탄소 중독", "작업 전 TBM 전파", 1, 3, "V", "1. 주기적인 환기(일 2회 이상) 실시\n2. 내부 화기 취급 금지 및 필요시 가스 감지기 설치"]
]

DATA_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "위험성평가_기본데이터.xlsx")

def init_data_file():
    if os.path.exists(DATA_FILE_PATH):
        return
    
    os.makedirs(os.path.dirname(DATA_FILE_PATH), exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    default_data = {
        "RT": data_rt,
        "UT": data_ut,
        "PT": data_pt,
        "컨테이너": data_container
    }
    
    headers = ["중공종", "위험요인(위험성평가)", "기존대책(미사용)", "빈도", "강도", "기존등급(미사용)", "위험성평가 개선대책"]
    
    for sheet_name, data in default_data.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for c in range(1, 8):
            ws.cell(row=1, column=c).font = Font(bold=True)
            ws.cell(row=1, column=c).fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            
        for row in data:
            ws.append(row)
            
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 80
            
    wb.save(DATA_FILE_PATH)

def load_data(sheet_name):
    init_data_file()
    wb = openpyxl.load_workbook(DATA_FILE_PATH, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
        
    ws = wb[sheet_name]
    data = []
    last_cat = ""
    last_risk = ""
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        
        # 원본 데이터 엑셀에 셀 병합이 되어 있거나 공백이 포함된 경우를 처리
        raw_cat = str(row[0]).strip() if row[0] is not None else ""
        raw_risk = str(row[1]).strip() if row[1] is not None else ""
        
        current_cat = raw_cat if raw_cat != "" else last_cat
        current_risk = raw_risk if raw_risk != "" else last_risk
        
        if not current_cat and not current_risk: continue
        
        last_cat = current_cat
        last_risk = current_risk
        
        item = [
            current_cat,
            current_risk,
            row[2] or "",
            row[3] if row[3] is not None else 1,
            row[4] if row[4] is not None else 1,
            row[5] or "",
            row[6] or ""
        ]
        data.append(item)
        
    # 세부작업(item[0]) 기준으로 논리적 순서에 맞게 정렬 (같은 작업끼리 묶이도록)
    order_map = {
        "이동": 1,
        "반입 및 설치": 2,
        "사전준비": 3,
        "비파괴검사": 4,
        "초음파탐상검사": 4,
        "침투탐상검사": 4,
        "전기설비": 5,
        "유지 및 운영": 6,
        "작업 후 정리": 7
    }
    data.sort(key=lambda x: (order_map.get(str(x[0]).strip(), 99), str(x[0])))
    
    return data

class RiskAssessmentApp:
    def __init__(self, root):
        self.root = root
        try:
            self.root.title("위험성 평가표 자동 생성기")
            self.root.geometry("500x720")
            self.root.resizable(True, True)
        except AttributeError:
            pass
        
        style = ttk.Style()
        style.theme_use('clam')
        
        main_frame = ttk.Frame(self.root, padding=20)
        main_frame.pack(fill='both', expand=True)
        
        # Title
        ttk.Label(main_frame, text="위험성 평가표 자동 생성기", font=('Malgun Gothic', 16, 'bold')).pack(pady=(0, 20))
        
        # Form Frame
        form_frame = ttk.LabelFrame(main_frame, text="기본 정보 설정", padding=15)
        form_frame.pack(fill='x', pady=5)
        
        config = load_config()
        
        # 1. 현장명
        ttk.Label(form_frame, text="현 장 명:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.ent_site = ttk.Entry(form_frame, width=45)
        self.ent_site.insert(0, config.get("site_name", "가산~가평 천연가스 공급시설 건설공사"))
        self.ent_site.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        
        # 2. 협력업체
        ttk.Label(form_frame, text="협력업체:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.ent_company = ttk.Entry(form_frame, width=45)
        self.ent_company.insert(0, config.get("company_name", "서울검사(주)"))
        self.ent_company.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        
        # 3. 작성일자
        ttk.Label(form_frame, text="작성일자:").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        self.ent_write_date = DateEntry(form_frame, width=42, date_pattern='yyyy년 mm월 dd일')
        # date pattern in DateEntry uses lowercase y, m, d
        # but DateEntry might not parse existing values properly if they have text in them.
        self.ent_write_date.delete(0, tk.END)
        self.ent_write_date.insert(0, config.get("write_date", datetime.now().strftime("%Y년 %m월 %d일")))
        self.ent_write_date.grid(row=2, column=1, sticky='w', padx=5, pady=5)
        
        # 4. 관리기간
        ttk.Label(form_frame, text="시작일자:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        self.ent_start_date = DateEntry(form_frame, width=42, date_pattern='yyyy년 mm월 dd일')
        self.ent_start_date.delete(0, tk.END)
        self.ent_start_date.insert(0, config.get("start_date", "2026년 06월 16일"))
        self.ent_start_date.grid(row=3, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(form_frame, text="종료일자:").grid(row=4, column=0, sticky='e', padx=5, pady=5)
        self.ent_end_date = DateEntry(form_frame, width=42, date_pattern='yyyy년 mm월 dd일')
        self.ent_end_date.delete(0, tk.END)
        self.ent_end_date.insert(0, config.get("end_date", "2026년 06월 30일"))
        self.ent_end_date.grid(row=4, column=1, sticky='w', padx=5, pady=5)
        
        
        # 6. 평가구분
        ttk.Label(form_frame, text="평가구분:").grid(row=6, column=0, sticky='e', padx=5, pady=5)
        self.cb_eval_type = ttk.Combobox(form_frame, values=["최초평가", "수시평가", "정기평가"], state='readonly', width=42)
        self.cb_eval_type.set(config.get("eval_type", "최초평가"))
        self.cb_eval_type.grid(row=6, column=1, sticky='w', padx=5, pady=5)
        
        
        # 8. 결재자 이름 설정
        ttk.Label(form_frame, text="수급인 근로자:").grid(row=8, column=0, sticky='e', padx=5, pady=5)
        self.ent_worker_name = ttk.Entry(form_frame, width=45)
        self.ent_worker_name.insert(0, config.get("worker_name", "유상훈"))
        self.ent_worker_name.grid(row=8, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(form_frame, text="수급인 작성자:").grid(row=9, column=0, sticky='e', padx=5, pady=5)
        self.ent_writer_name = ttk.Entry(form_frame, width=45)
        self.ent_writer_name.insert(0, config.get("writer_name", "주진철"))
        self.ent_writer_name.grid(row=9, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(form_frame, text="수급인 승인자:").grid(row=10, column=0, sticky='e', padx=5, pady=5)
        self.ent_approver_name = ttk.Entry(form_frame, width=45)
        self.ent_approver_name.insert(0, config.get("approver_name", "강신태"))
        self.ent_approver_name.grid(row=10, column=1, sticky='w', padx=5, pady=5)
        
        # Checkbox Frame for selections
        chk_frame = ttk.LabelFrame(main_frame, text="생성 대상 선택", padding=15)
        chk_frame.pack(fill='x', pady=10)
        
        self.var_rt = tk.BooleanVar(value=True)
        self.var_ut = tk.BooleanVar(value=True)
        self.var_pt = tk.BooleanVar(value=True)
        self.var_container = tk.BooleanVar(value=True)
        
        ttk.Checkbutton(chk_frame, text="RT", variable=self.var_rt).grid(row=0, column=0, padx=10, pady=5, sticky='w')
        ttk.Checkbutton(chk_frame, text="UT", variable=self.var_ut).grid(row=0, column=1, padx=10, pady=5, sticky='w')
        ttk.Checkbutton(chk_frame, text="PT", variable=self.var_pt).grid(row=0, column=2, padx=10, pady=5, sticky='w')
        ttk.Checkbutton(chk_frame, text="컨테이너", variable=self.var_container).grid(row=0, column=3, padx=10, pady=5, sticky='w')
        
        # Generate Button
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=10)
        
        self.btn_edit_data = ttk.Button(btn_frame, text="기본 데이터 수정하기 (엑셀 연동)", command=self.open_data_file, width=35)
        self.btn_edit_data.pack(pady=(0, 5))
        
        self.btn_generate = ttk.Button(btn_frame, text="선택한 위험성 평가표 일괄 생성", command=self.generate_files, width=35)
        self.btn_generate.pack(pady=5)
        
        self.lbl_status = ttk.Label(main_frame, text="대기 중...", foreground="gray")
        self.lbl_status.pack()

    def open_data_file(self):
        init_data_file()
        try:
            os.startfile(DATA_FILE_PATH)
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 파일을 여는 중 오류가 발생했습니다:\n{e}")

    def generate_files(self, silent_dir=None, date_str=""):
        params = {
            'site_name': self.ent_site.get().strip(),
            'company_name': self.ent_company.get().strip(),
            'write_date': self.ent_write_date.get().strip(),
            'start_date': self.ent_start_date.get().strip(),
            'end_date': self.ent_end_date.get().strip(),
            'eval_type': self.cb_eval_type.get(),
            'worker_name': self.ent_worker_name.get().strip(),
            'writer_name': self.ent_writer_name.get().strip(),
            'approver_name': self.ent_approver_name.get().strip()
        }
        
        config = load_config()
        config.update(params)
        
        if silent_dir:
            output_dir = silent_dir
        else:
            initial_dir = config.get("last_output_dir", os.path.dirname(os.path.abspath(__file__)))
            if not os.path.exists(initial_dir):
                initial_dir = os.path.dirname(os.path.abspath(__file__))
                
            output_dir = filedialog.askdirectory(title="저장할 폴더를 선택하세요", initialdir=initial_dir)
            if not output_dir:
                return
                
            config["last_output_dir"] = output_dir
            
        save_config(config)
            
        self.btn_generate.config(state='disabled')
        self.lbl_status.config(text="생성 중...", foreground="blue")
        self.root.update()
        
        results = []
        temp_files = []
        
        try:
            prefix_idx = 1
            if self.var_rt.get():
                fname = os.path.join(output_dir, f"temp_risk_rt_{date_str}.xlsx") if silent_dir else os.path.join(output_dir, f"4.4.{prefix_idx}_위험성평가표(RT_표준양식).xlsx")
                res, msg = create_excel("방사선투과검사", fname, load_data("RT"), params)
                if res: 
                    results.append(msg)
                    temp_files.append(fname)
                    prefix_idx += 1
                
            if self.var_ut.get():
                fname = os.path.join(output_dir, f"temp_risk_ut_{date_str}.xlsx") if silent_dir else os.path.join(output_dir, f"4.4.{prefix_idx}_위험성평가표(UT_표준양식).xlsx")
                res, msg = create_excel("초음파탐상검사", fname, load_data("UT"), params)
                if res: 
                    results.append(msg)
                    temp_files.append(fname)
                    prefix_idx += 1
                
            if self.var_pt.get():
                fname = os.path.join(output_dir, f"temp_risk_pt_{date_str}.xlsx") if silent_dir else os.path.join(output_dir, f"4.4.{prefix_idx}_위험성평가표(PT_표준양식).xlsx")
                res, msg = create_excel("침투탐상검사", fname, load_data("PT"), params)
                if res: 
                    results.append(msg)
                    temp_files.append(fname)
                    prefix_idx += 1
                
            if self.var_container.get():
                fname = os.path.join(output_dir, f"temp_risk_ct_{date_str}.xlsx") if silent_dir else os.path.join(output_dir, f"4.4.{prefix_idx}_위험성평가표(컨테이너_표준양식).xlsx")
                res, msg = create_excel("가설컨테이너 설치 및 운영", fname, load_data("컨테이너"), params)
                if res: 
                    results.append(msg)
                    temp_files.append(fname)
                    prefix_idx += 1
                
            if silent_dir:
                return temp_files
                
            if results:
                messagebox.showinfo("생성 완료", "\n".join(results))
                self.lbl_status.config(text="생성 완료!", foreground="green")
            else:
                messagebox.showwarning("경고", "생성할 항목을 최소 하나 이상 선택해주세요.")
                self.lbl_status.config(text="항목 미선택", foreground="red")
                
        except Exception as e:
            if not silent_dir:
                messagebox.showerror("오류", f"엑셀 생성 중 오류가 발생했습니다:\n{str(e)}")
            self.lbl_status.config(text="오류 발생", foreground="red")
        finally:
            self.btn_generate.config(state='normal')
            
        if silent_dir:
            return temp_files

if __name__ == "__main__":
    root = tk.Tk()
    app = RiskAssessmentApp(root)
    root.mainloop()
