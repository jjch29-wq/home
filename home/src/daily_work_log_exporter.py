import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
import os

class DailyWorkLogExporter:
    def __init__(self):
        # Define basic styles
        self.font_title = Font(name='맑은 고딕', size=16, bold=True)
        self.font_normal = Font(name='맑은 고딕', size=9)
        self.font_bold = Font(name='맑은 고딕', size=9, bold=True)
        self.font_small = Font(name='맑은 고딕', size=8)
        
        self.align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        thin = Side(style='thin')
        self.border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        self.fill_header = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    def generate_excel(self, data, output_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "작업감독일보"

        # 1. Set Column Widths (A to AA roughly)
        col_widths = {
            'A': 4, 'B': 14, 'C': 7, 'D': 8, 'E': 8, 'F': 8, 'G': 8, 'H': 6, 'I': 7, 'J': 6, 
            'K': 2, 'L': 9, 'M': 5, 'N': 5, 'O': 9, 'P': 9, 'Q': 2, 'R': 4, 'S': 4, 'T': 4, 'U': 4, 'V': 4, 'W': 4, 'X': 4, 'Y': 4, 'Z': 4
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Helper to set cell value and style
        def set_cell(coord, value, font=self.font_normal, align=self.align_center, border=self.border_thin, fill=None):
            cell = ws[coord]
            cell.value = value
            cell.font = font
            cell.alignment = align
            if border: cell.border = border
            if fill: cell.fill = fill
            return cell

        def merge_and_set(range_str, value, font=self.font_normal, align=self.align_center, border=self.border_thin, fill=None):
            ws.merge_cells(range_str)
            start_cell = range_str.split(':')[0]
            cell = set_cell(start_cell, value, font, align, None, fill)
            # Apply border to all cells in merged range to avoid missing lines
            for row in ws[range_str]:
                for c in row:
                    if border: c.border = border
            return cell

        # =========================================================================
        # HEADER SECTION
        # =========================================================================
        merge_and_set('A2:P2', "비파괴검사 결과서 및 작업/감독일보", font=self.font_title, border=None)
        
        # Left Info
        merge_and_set('A4:F4', "용 역 명 : 2026년 중앙지사 열수송관 비파괴검사용역", font=self.font_normal, align=self.align_left, border=None)
        date_str = data.get('date', '2026년   월   일')
        weather_str = data.get('weather', '')
        merge_and_set('A5:F5', f"검사일자 : {date_str}          날씨 : {weather_str}", font=self.font_normal, align=self.align_left, border=None)

        # Right Info (Signatures)
        merge_and_set('L4:M4', f"Page {data.get('page_current', 1)} of {data.get('page_total', 1)}", font=self.font_normal, align=self.align_right, border=None)
        
        set_cell('N4', '현장대리인', font=self.font_small, fill=self.fill_header)
        set_cell('O4', '감독', font=self.font_small, fill=self.fill_header)
        set_cell('P4', '확인', font=self.font_small, fill=self.fill_header)
        
        set_cell('N5', '') # Signature space
        set_cell('O5', '')
        set_cell('P5', '')
        ws.row_dimensions[5].height = 30 # Make signature box taller

        # =========================================================================
        # SECTION 1: 작업 물량 및 누계 현황 (A7:J22)
        # =========================================================================
        merge_and_set('A7:F7', "1. 작업 물량 및 누계 현황", font=self.font_bold, align=self.align_left, border=None)
        
        headers_qty = ['방법', '규격', '예상량', '전일 누계', '금일 작업', '총 누계', '공정률(%)', '불량', '불량률(%)', '비고']
        for i, header in enumerate(headers_qty):
            col_letter = get_column_letter(i+1)
            set_cell(f'{col_letter}8', header, font=self.font_bold, fill=self.fill_header)

        qty_rows = [
            ('PAUT', '300A이상'), ('PAUT', '300A이상-야간'), ('PAUT', '250A'), ('PAUT', '200A'), ('PAUT', '200A-야간'), ('PAUT', '소계'),
            ('RT', '150A~100A'), ('RT', '150A~100A-야간'), ('RT', '80A이하'), ('RT', '80A이하-야간'), ('RT', '소계'),
            ('MT', '전체(주간)'), ('MT', '전체(야간)'),
            ('PT', '전체(주간)'), ('PT', '전체(야간)')
        ]
        
        row_idx = 9
        for method, spec in qty_rows:
            set_cell(f'A{row_idx}', method)
            set_cell(f'B{row_idx}', spec)
            
            # Fetch data from input dictionary
            row_data = data.get('qty_data', {}).get(spec, {})
            set_cell(f'C{row_idx}', row_data.get('예상량', ''))
            set_cell(f'D{row_idx}', row_data.get('전일누계', ''))
            set_cell(f'E{row_idx}', row_data.get('금일작업', ''))
            set_cell(f'F{row_idx}', row_data.get('총누계', ''))
            set_cell(f'G{row_idx}', row_data.get('공정률', ''))
            set_cell(f'H{row_idx}', row_data.get('불량', ''))
            set_cell(f'I{row_idx}', row_data.get('불량률', ''))
            set_cell(f'J{row_idx}', row_data.get('비고', ''))
            
            row_idx += 1
            
        ws.merge_cells('A9:A14') # PAUT
        ws.merge_cells('A15:A19') # RT
        ws.merge_cells('A20:A21') # MT
        ws.merge_cells('A22:A23') # PT

        # =========================================================================
        # RIGHT SIDE SECTION: Equipment and Personnel (L7:P23)
        # =========================================================================
        # Equipment
        merge_and_set('L8:N8', '장비투입 현황(대)', font=self.font_bold, fill=self.fill_header)
        set_cell('L9', '장비명', font=self.font_bold, fill=self.fill_header)
        set_cell('M9', '금일', font=self.font_bold, fill=self.fill_header)
        set_cell('N9', '누계', font=self.font_bold, fill=self.fill_header)
        
        equip_rows = ['PAUT장비', 'PAUT프로브', 'PAUT스캐너', 'RT장비', 'MT장비']
        for i, eq in enumerate(equip_rows):
            r = 10 + i
            set_cell(f'L{r}', eq)
            eq_data = data.get('equip_data', {}).get(eq, {})
            set_cell(f'M{r}', eq_data.get('금일', ''))
            set_cell(f'N{r}', eq_data.get('누계', ''))

        # Personnel
        merge_and_set('O8:P8', '금일 투입 인원현황(명)', font=self.font_bold, fill=self.fill_header)
        set_cell('O9', '구분 (관리/안전)', font=self.font_small, fill=self.fill_header)
        set_cell('P9', '검사원', font=self.font_bold, fill=self.fill_header)
        
        personnel = data.get('personnel_data', {})
        set_cell('O10', '인원')
        set_cell('P10', personnel.get('검사원_인원', ''))
        set_cell('O11', '현장대리인')
        set_cell('P11', personnel.get('검사원_현장대리인', ''))
        set_cell('O12', '누계')
        set_cell('P12', personnel.get('검사원_누계', ''))
        
        # Merge empty blocks under personnel to match equipment height
        merge_and_set('O13:P14', '', border=self.border_thin)

        # Remarks (특이사항 및 작업계획)
        merge_and_set('L15:P15', '특이사항 및 작업계획', font=self.font_bold, fill=self.fill_header)
        merge_and_set('L16:P23', data.get('remarks', ''), align=Alignment(horizontal='left', vertical='top', wrap_text=True))

        # =========================================================================
        # SECTION 2: 비파괴검사결과서 (A26:Y...)
        # =========================================================================
        merge_and_set('A25:F25', "2. 비파괴검사결과서", font=self.font_bold, align=self.align_left, border=None)
        
        # Main Headers
        headers_ndt = [
            ('A26:A27', '순번'), ('B26:B27', '검사방법'), ('C26:C27', '구간(Section No.)'), ('D26:D27', '라인번호'),
            ('E26:E27', 'Joint No.'), ('F26:F27', '관경'), ('G26:G27', '용접사'), ('H26:H27', '구간정보(Start/Length)'),
            ('I26:I27', '결과'), ('J26:J27', '규격')
        ]
        for rng, text in headers_ndt:
            merge_and_set(rng, text, font=self.font_small, fill=self.fill_header)
            
        merge_and_set('K26:L26', 'RT매수', font=self.font_small, fill=self.fill_header)
        set_cell('K27', 'OR', font=self.font_small, fill=self.fill_header)
        set_cell('L27', 'RE', font=self.font_small, fill=self.fill_header)
        
        merge_and_set('M26:O26', 'PAUT길이(m)', font=self.font_small, fill=self.fill_header)
        set_cell('M27', '주간', font=self.font_small, fill=self.fill_header)
        set_cell('N27', '야간', font=self.font_small, fill=self.fill_header)
        set_cell('O27', '재검', font=self.font_small, fill=self.fill_header)
        
        merge_and_set('P26:Q26', 'MT(m)', font=self.font_small, fill=self.fill_header)
        set_cell('P27', '주간', font=self.font_small, fill=self.fill_header)
        set_cell('Q27', '야간', font=self.font_small, fill=self.fill_header)
        
        merge_and_set('R26:S26', 'PT(m)', font=self.font_small, fill=self.fill_header)
        set_cell('R27', '주간', font=self.font_small, fill=self.fill_header)
        set_cell('S27', '야간', font=self.font_small, fill=self.fill_header)
        
        # Populate NDT Results
        ndt_results = data.get('ndt_results', [])
        start_row = 28
        for i in range(max(15, len(ndt_results))): # minimum 15 empty rows for good look
            row_idx = start_row + i
            res = ndt_results[i] if i < len(ndt_results) else {}
            
            set_cell(f'A{row_idx}', i + 1)
            set_cell(f'B{row_idx}', res.get('검사방법', ''))
            set_cell(f'C{row_idx}', res.get('구간', ''))
            set_cell(f'D{row_idx}', res.get('라인번호', ''))
            set_cell(f'E{row_idx}', res.get('Joint No.', ''))
            set_cell(f'F{row_idx}', res.get('관경', ''))
            set_cell(f'G{row_idx}', res.get('용접사', ''))
            set_cell(f'H{row_idx}', res.get('구간정보', ''))
            set_cell(f'I{row_idx}', res.get('결과', ''))
            set_cell(f'J{row_idx}', res.get('규격', ''))
            set_cell(f'K{row_idx}', res.get('RT_OR', ''))
            set_cell(f'L{row_idx}', res.get('RT_RE', ''))
            set_cell(f'M{row_idx}', res.get('PAUT_주간', ''))
            set_cell(f'N{row_idx}', res.get('PAUT_야간', ''))
            set_cell(f'O{row_idx}', res.get('PAUT_재검', ''))
            set_cell(f'P{row_idx}', res.get('MT_주간', ''))
            set_cell(f'Q{row_idx}', res.get('MT_야간', ''))
            set_cell(f'R{row_idx}', res.get('PT_주간', ''))
            set_cell(f'S{row_idx}', res.get('PT_야간', ''))
            
            ws.row_dimensions[row_idx].height = 20

        # Print setup (Landscape, fit to one page width)
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        
        # Margins
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        wb.save(output_path)
        return output_path

if __name__ == "__main__":
    # Test generation
    exporter = DailyWorkLogExporter()
    test_data = {
        'date': '2026년 08월 05일',
        'weather': '맑음',
        'qty_data': {
            '300A이상': {'예상량': 129, '전일누계': 100, '금일작업': 10, '총누계': 110, '공정률': '85.2%'}
        },
        'remarks': '안전 작업 지시 이행\nPAUT 장비 점검 완료',
        'ndt_results': [
            {'검사방법': 'PAUT', '구간': 'S-1', '라인번호': 'L-1', 'Joint No.': 'J-1', '결과': '합격'}
        ]
    }
    # exporter.generate_excel(test_data, "test_daily_report.xlsx")
