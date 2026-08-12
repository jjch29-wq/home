"""PAUT Scan Plan Excel generator with a small Tkinter desktop UI."""

from __future__ import annotations

import os
import re
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageTk


APP_TITLE = "PAUT Scan Plan 엑셀 생성기"


@dataclass(frozen=True)
class ScanPlanConfig:
    image_path: Path
    output_path: Path
    thickness: float
    probe: str = "5L64-A2"
    min_angle: float = 46.0
    max_angle: float = 61.0
    angle_step: float = 1.0
    indexes: tuple[float, float, float] = (22.0, 44.0, 67.0)
    scan_direction: str = "양측 주사"
    active_elements: int = 32
    first_element: int = 1
    focal_length: float = 11.0


PROBES = {
    "5L64-A2": {
        "frequency": 5.0,
        "elements": 64,
        "pitch": 0.60,
        "total_aperture": "38.4×10",
        "elevation": 10.0,
    }
}


def _fmt(value: float) -> str:
    return f"{value:g}"


def default_output_path(image_path: Path, thickness: float, min_angle: float, max_angle: float) -> Path:
    name = f"PAUT_검사설정_5L64-A2_{_fmt(thickness)}mm_{_fmt(min_angle)}-{_fmt(max_angle)}도.xlsx"
    return image_path.parent / name


def validate_config(config: ScanPlanConfig) -> None:
    if not config.image_path.is_file():
        raise ValueError("Scan Plan 이미지 파일을 선택해 주세요.")
    if config.image_path.suffix.lower() not in {".png", ".jpg", ".jpeg", ".bmp"}:
        raise ValueError("PNG, JPG, JPEG 또는 BMP 이미지만 사용할 수 있습니다.")
    if config.probe not in PROBES:
        raise ValueError(f"지원하지 않는 탐촉자입니다: {config.probe}")
    if config.thickness <= 0:
        raise ValueError("부재 두께는 0보다 커야 합니다.")
    if not 0 < config.min_angle < config.max_angle < 90:
        raise ValueError("빔각은 0°~90° 사이이며 최소각이 최대각보다 작아야 합니다.")
    if config.angle_step <= 0:
        raise ValueError("Angle Step은 0보다 커야 합니다.")
    if any(value < 0 for value in config.indexes):
        raise ValueError("Index Offset은 0 이상이어야 합니다.")
    if config.active_elements <= 0 or config.active_elements > PROBES[config.probe]["elements"]:
        raise ValueError("Active Element 수가 탐촉자의 전체 진동자 수 범위를 벗어났습니다.")
    if config.first_element < 1:
        raise ValueError("First Element는 1 이상이어야 합니다.")


def create_workbook(config: ScanPlanConfig) -> Path:
    validate_config(config)
    spec = PROBES[config.probe]
    output = config.output_path.with_suffix(".xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "PAUT 설정표"
    ws.sheet_view.showGridLines = False

    navy, blue, yellow, white = "17365D", "D9EAF7", "FFF2CC", "FFFFFF"
    thin = Side(style="thin", color="808080")
    medium = Side(style="medium", color="000000")

    def style_range(cell_range: str, fill: str, *, bold=False, color="000000", border=thin, wrap=True):
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.font = Font(name="맑은 고딕", bold=bold, color=color, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
                cell.border = Border(left=border, right=border, top=border, bottom=border)

    ws.merge_cells("A1:K2")
    ws["A1"] = "PAUT 검사 설정표"
    ws["A1"].font = Font(name="맑은 고딕", size=18, bold=True, color=white)
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A4:K4")
    ws["A4"] = "1. 기본 입력조건"
    style_range("A4:K4", navy, bold=True, color=white, border=medium)

    inputs = [
        ("A5", "부재 두께 (mm)", "B5", config.thickness),
        ("D5", "공칭 개선각 (°)", "E5", 37.5),
        ("G5", "개선각 공차 (±°)", "H5", 2.5),
        ("J5", "Angle step (°)", "K5", config.angle_step),
        ("A6", "최소 빔각 (°)", "B6", config.min_angle),
        ("D6", "최대 빔각 (°)", "E6", config.max_angle),
        ("G6", "검사 방향", "H6", config.scan_direction),
        ("J6", "Wave type", "K6", "SW"),
    ]
    for label_cell, label, value_cell, value in inputs:
        ws[label_cell], ws[value_cell] = label, value
        style_range(f"{label_cell}:{label_cell}", blue, bold=True)
        style_range(f"{value_cell}:{value_cell}", yellow)
        ws[value_cell].font = Font(name="맑은 고딕", size=9, color="0000FF")

    main_headers = ["Probe", "Wave type", "Law config.", "Focus type", "Aperture Area", "Sweep angle range", "Angle resolution", "Focal depth"]
    for col, text in enumerate(main_headers, 1):
        ws.cell(8, col, text)
    ws.merge_cells("I8:K8")
    ws["I8"] = "Index offset (mm)"
    style_range("A8:K8", blue, bold=True, border=medium)

    for col, value in zip(range(9, 12), config.indexes):
        ws.cell(9, col, value)
    active_width = config.active_elements * spec["pitch"]
    main_values = [
        "1D", "SW", "Sectorial", "True depth",
        f"{config.probe}\n{config.active_elements} elements × {spec['pitch']:.2f} mm\nActive: {active_width:.1f}×{spec['elevation']:g} mm",
        f"{_fmt(config.min_angle)}°~{_fmt(config.max_angle)}°\n{_fmt(config.angle_step)}° increment",
        config.angle_step, "Bottom (ID)",
    ]
    for col, value in enumerate(main_values, 1):
        ws.cell(9, col, value)
        ws.merge_cells(start_row=9, start_column=col, end_row=10, end_column=col)
    ws["I10"], ws["J10"], ws["K10"] = "Direct / near side", "Reflected path", "Full-skip check"
    style_range("A9:K10", white, bold=True, border=medium)
    ws.row_dimensions[9].height = 48
    ws.row_dimensions[10].height = 27

    ws.merge_cells("A12:K12")
    ws["A12"] = "2. 탐촉자 정보"
    style_range("A12:K12", navy, bold=True, color=white, border=medium)
    probe_headers = ["탐촉자 형식", "주파수 (MHz)", "진동자 수", "전체 개구 (mm)", "Pitch (mm)", "Element Qty", "First Element", "Focusing Type", "Focal Length", "Index", "비고"]
    for col, text in enumerate(probe_headers, 1):
        ws.cell(13, col, text)
    style_range("A13:K13", blue, bold=True, border=medium)
    probe_values = [
        config.probe, spec["frequency"], spec["elements"], spec["total_aperture"], spec["pitch"],
        config.active_elements, config.first_element, "TRUE", config.focal_length,
        " / ".join(_fmt(v) for v in config.indexes), f"Active aperture: {active_width:.1f}×{spec['elevation']:g} mm",
    ]
    for col, value in enumerate(probe_values, 1):
        ws.cell(14, col, value)
    style_range("A14:K14", white)
    ws.row_dimensions[14].height = 30

    ws.merge_cells("A16:K16")
    ws["A16"] = (
        f"3. PAUT Scan Plan (두께 {_fmt(config.thickness)} mm / 빔각 {_fmt(config.min_angle)}°~{_fmt(config.max_angle)}° / "
        f"Index {'·'.join(_fmt(v) for v in config.indexes)} mm / {config.scan_direction})"
    )
    style_range("A16:K16", navy, bold=True, color=white, border=medium)

    with Image.open(config.image_path) as source_image:
        iw, ih = source_image.size
    target_width = 1240
    target_height = int(target_width * ih / iw)
    if target_height > 430:
        target_height = 430
        target_width = int(target_height * iw / ih)
    picture = XLImage(str(config.image_path))
    picture.width, picture.height = target_width, target_height
    picture.anchor = "B17"
    ws.add_image(picture)
    for row in range(17, 35):
        ws.row_dimensions[row].height = 17
    ws.merge_cells("A35:K35")
    ws["A35"] = "Scan Plan 프로그램 출력값이며 실제 웨지 출사점·용접 형상·덧살 및 감도는 Demonstration Block으로 확인한다."
    ws["A35"].font = Font(name="맑은 고딕", size=9, italic=True, color="666666")
    ws["A35"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [18, 15, 16, 18, 22, 20, 18, 18, 17, 30, 25]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.print_area = "A1:K35"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.25
    ws.page_margins.top = ws.page_margins.bottom = 0.30
    ws.page_margins.header = ws.page_margins.footer = 0.15
    ws.oddFooter.center.text = "PAUT 검사 설정표"
    ws.oddFooter.right.text = "Page &P of &N"
    ws.freeze_panes = "A8"
    ws.sheet_view.zoomScale = 75
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output)
    return output


def verify_workbook(path: Path, config: ScanPlanConfig) -> None:
    wb = load_workbook(path, data_only=False)
    ws = wb["PAUT 설정표"]
    if ws["B5"].value != config.thickness or ws["A14"].value != config.probe:
        raise RuntimeError("생성된 파일의 기본 설정 검증에 실패했습니다.")
    if [ws[cell].value for cell in ("I9", "J9", "K9")] != list(config.indexes):
        raise RuntimeError("생성된 파일의 Index Offset 검증에 실패했습니다.")
    if len(ws._images) != 1:
        raise RuntimeError("Scan Plan 그림 삽입 검증에 실패했습니다.")
    if ws.page_setup.fitToWidth != 1 or ws.page_setup.fitToHeight != 1:
        raise RuntimeError("A4 1페이지 인쇄 설정 검증에 실패했습니다.")


def open_file(path: Path) -> None:
    if sys.platform == "win32":
        os.startfile(path)  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(path)])
    else:
        subprocess.Popen(["xdg-open", str(path)])


class ScanPlanGeneratorApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("900x720")
        self.minsize(820, 650)
        self.last_output: Path | None = None
        self.preview_photo: ImageTk.PhotoImage | None = None
        self._configure_style()
        self._build_ui()

    def _configure_style(self) -> None:
        style = ttk.Style(self)
        if "vista" in style.theme_names():
            style.theme_use("vista")
        style.configure("Title.TLabel", font=("맑은 고딕", 17, "bold"), foreground="#17365D")
        style.configure("Section.TLabelframe.Label", font=("맑은 고딕", 10, "bold"), foreground="#17365D")
        style.configure("Primary.TButton", font=("맑은 고딕", 10, "bold"))

    def _build_ui(self) -> None:
        root = ttk.Frame(self, padding=16)
        root.pack(fill="both", expand=True)
        ttk.Label(root, text=APP_TITLE, style="Title.TLabel").pack(anchor="w", pady=(0, 12))

        form = ttk.LabelFrame(root, text="입력조건", style="Section.TLabelframe", padding=12)
        form.pack(fill="x")
        form.columnconfigure(1, weight=1)

        self.image_var = tk.StringVar()
        self.thickness_var = tk.StringVar(value="25.47")
        self.probe_var = tk.StringVar(value="5L64-A2")
        self.min_angle_var = tk.StringVar(value="46")
        self.max_angle_var = tk.StringVar(value="61")
        self.step_var = tk.StringVar(value="1")
        self.index_vars = [tk.StringVar(value=v) for v in ("22", "44", "67")]
        self.direction_var = tk.StringVar(value="양측 주사")
        self.output_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Scan Plan 이미지를 선택해 주세요.")

        ttk.Label(form, text="Scan Plan 이미지").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=5)
        ttk.Entry(form, textvariable=self.image_var).grid(row=0, column=1, columnspan=5, sticky="ew", pady=5)
        ttk.Button(form, text="파일 선택", command=self.choose_image).grid(row=0, column=6, padx=(8, 0), pady=5)

        fields = [
            ("부재 두께 (mm)", self.thickness_var), ("탐촉자", self.probe_var),
            ("최소 빔각 (°)", self.min_angle_var), ("최대 빔각 (°)", self.max_angle_var),
            ("Angle Step (°)", self.step_var), ("검사 방향", self.direction_var),
        ]
        for idx, (label, variable) in enumerate(fields):
            row, offset = 1 + idx // 3, (idx % 3) * 2
            ttk.Label(form, text=label).grid(row=row, column=offset, sticky="w", padx=(0, 6), pady=5)
            if label == "탐촉자":
                widget = ttk.Combobox(form, textvariable=variable, values=list(PROBES), state="readonly", width=15)
            elif label == "검사 방향":
                widget = ttk.Combobox(form, textvariable=variable, values=["양측 주사", "단측 주사"], state="readonly", width=15)
            else:
                widget = ttk.Entry(form, textvariable=variable, width=15)
            widget.grid(row=row, column=offset + 1, sticky="ew", padx=(0, 12), pady=5)

        ttk.Label(form, text="Index Offset (mm)").grid(row=3, column=0, sticky="w", pady=5)
        index_frame = ttk.Frame(form)
        index_frame.grid(row=3, column=1, columnspan=5, sticky="w", pady=5)
        for i, variable in enumerate(self.index_vars):
            ttk.Entry(index_frame, textvariable=variable, width=10).pack(side="left", padx=(0, 8))

        ttk.Label(form, text="출력 파일").grid(row=4, column=0, sticky="w", padx=(0, 8), pady=5)
        ttk.Entry(form, textvariable=self.output_var).grid(row=4, column=1, columnspan=5, sticky="ew", pady=5)
        ttk.Button(form, text="저장 위치", command=self.choose_output).grid(row=4, column=6, padx=(8, 0), pady=5)

        preview_box = ttk.LabelFrame(root, text="이미지 미리보기", style="Section.TLabelframe", padding=8)
        preview_box.pack(fill="both", expand=True, pady=12)
        self.preview_label = ttk.Label(preview_box, text="선택된 이미지가 없습니다.", anchor="center")
        self.preview_label.pack(fill="both", expand=True)

        actions = ttk.Frame(root)
        actions.pack(fill="x")
        ttk.Label(actions, textvariable=self.status_var).pack(side="left", fill="x", expand=True)
        ttk.Button(actions, text="생성 파일 열기", command=self.open_last_output).pack(side="right", padx=(8, 0))
        ttk.Button(actions, text="엑셀 생성", style="Primary.TButton", command=self.generate).pack(side="right")

    def choose_image(self) -> None:
        selected = filedialog.askopenfilename(
            title="Scan Plan 이미지 선택",
            filetypes=[("이미지", "*.png *.jpg *.jpeg *.bmp"), ("모든 파일", "*.*")],
        )
        if not selected:
            return
        self.image_var.set(selected)
        self._show_preview(Path(selected))
        try:
            thickness = float(self.thickness_var.get())
            self.output_var.set(str(default_output_path(Path(selected), thickness, float(self.min_angle_var.get()), float(self.max_angle_var.get()))))
        except ValueError:
            pass
        self.status_var.set("이미지를 선택했습니다. 그림의 두께와 Index 값을 확인해 입력하세요.")

    def _show_preview(self, path: Path) -> None:
        try:
            with Image.open(path) as source:
                image = source.convert("RGB")
                image.thumbnail((820, 360), Image.Resampling.LANCZOS)
            self.preview_photo = ImageTk.PhotoImage(image)
            self.preview_label.configure(image=self.preview_photo, text="")
        except Exception as exc:
            self.preview_label.configure(image="", text=f"미리보기를 표시할 수 없습니다.\n{exc}")

    def choose_output(self) -> None:
        selected = filedialog.asksaveasfilename(
            title="출력 엑셀 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 통합문서", "*.xlsx")],
            initialfile=Path(self.output_var.get()).name if self.output_var.get() else "PAUT_검사설정.xlsx",
        )
        if selected:
            self.output_var.set(selected)

    def _config(self) -> ScanPlanConfig:
        image_path = Path(self.image_var.get().strip())
        thickness = float(self.thickness_var.get())
        min_angle, max_angle = float(self.min_angle_var.get()), float(self.max_angle_var.get())
        output_text = self.output_var.get().strip()
        output = Path(output_text) if output_text else default_output_path(image_path, thickness, min_angle, max_angle)
        return ScanPlanConfig(
            image_path=image_path,
            output_path=output,
            thickness=thickness,
            probe=self.probe_var.get(),
            min_angle=min_angle,
            max_angle=max_angle,
            angle_step=float(self.step_var.get()),
            indexes=tuple(float(v.get()) for v in self.index_vars),  # type: ignore[arg-type]
            scan_direction=self.direction_var.get(),
        )

    def generate(self) -> None:
        try:
            config = self._config()
            self.status_var.set("엑셀 파일을 생성하고 검증하는 중입니다...")
            self.update_idletasks()
            output = create_workbook(config)
            verify_workbook(output, config)
            self.last_output = output
            self.output_var.set(str(output))
            self.status_var.set(f"생성 완료: {output.name}")
            messagebox.showinfo(APP_TITLE, f"엑셀 파일을 생성했습니다.\n\n{output}")
        except PermissionError:
            messagebox.showerror(APP_TITLE, "출력 파일이 Excel에서 열려 있습니다. 파일을 닫거나 다른 이름으로 저장해 주세요.")
            self.status_var.set("저장 실패: 출력 파일이 사용 중입니다.")
        except (ValueError, OSError, RuntimeError) as exc:
            messagebox.showerror(APP_TITLE, str(exc))
            self.status_var.set(f"오류: {exc}")

    def open_last_output(self) -> None:
        path = self.last_output or (Path(self.output_var.get()) if self.output_var.get() else None)
        if not path or not path.exists():
            messagebox.showwarning(APP_TITLE, "먼저 엑셀 파일을 생성해 주세요.")
            return
        open_file(path)


def main() -> None:
    ScanPlanGeneratorApp().mainloop()


if __name__ == "__main__":
    main()
