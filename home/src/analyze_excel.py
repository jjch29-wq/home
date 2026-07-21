import openpyxl

OUTPUT_FILE = r"c:\Users\jjch2\Desktop\PMI\home\src\excel_output.txt"

wb = openpyxl.load_workbook(r"C:\Users\jjch2\Desktop\월용역진도보고서.xlsx", data_only=True)

lines = []
lines.append(f"=== 시트 목록 ({len(wb.sheetnames)}개) ===")
for i, s in enumerate(wb.sheetnames):
    lines.append(f"  {i+1}. {s}")

for sname in wb.sheetnames:
    ws = wb[sname]
    lines.append(f"\n{'='*70}")
    lines.append(f"시트: [{sname}]  행수: {ws.max_row}  열수: {ws.max_column}  병합셀: {len(ws.merged_cells.ranges)}개")
    lines.append(f"{'='*70}")
    
    max_show = min(ws.max_row, 60)
    for row_idx in range(1, max_show + 1):
        vals = []
        for col_idx in range(1, min(ws.max_column + 1, 25)):
            cell = ws.cell(row=row_idx, column=col_idx)
            v = cell.value
            if v is not None:
                v_str = str(v).replace('\xa0', ' ').replace('\n', '\\n').strip()
                if v_str:
                    vals.append(f"C{col_idx}={v_str}")
        if vals:
            lines.append(f"  R{row_idx}: {' | '.join(vals)}")
    
    if ws.max_row > max_show:
        lines.append(f"  ... ({ws.max_row - max_show}개 행 추가)")

wb.close()

with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print("Done")
