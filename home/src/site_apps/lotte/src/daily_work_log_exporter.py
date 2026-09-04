import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
import os

class DailyWorkLogExporter:
    def __init__(self):
        # Define basic styles
        self.font_title = Font(name='맑은 고딕', size=16, bold=True)
        self.font_normal = Font(name='맑은 고딕', size=9)
        self.font_bold = Font(name='맑은 고딕', size=9, bold=True)
        self.font_small = Font(name='맑은 고딕', size=8)
        
        self.align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
        self.align_nowrap = Alignment(horizontal='center', vertical='center', wrap_text=False, shrink_to_fit=True)
        
        thin = Side(style='thin')
        self.border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        self.fill_header = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    def generate_excel(self, data, output_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "작업감독일보"

        # 1. Set Column Widths (A to AA roughly)
        col_widths = {
            'A': 5, 'B': 12, 'C': 8, 'D': 14, 'E': 21, 'F': 9, 'G': 8, 'H': 8, 'I': 15, 'J': 7, 'K': 7, 'L': 7,
            'M': 7, 'N': 6, 'O': 10, 'P': 10, 'Q': 10, 'R': 7, 'S': 7, 'T': 7, 'U': 4, 'V': 4, 'W': 4, 'X': 4, 'Y': 4, 'Z': 4
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Helper to set cell value and style
        def set_cell(coord, value, font=self.font_normal, align=self.align_center, border=self.border_thin, fill=None):
            cell = ws[coord]
            cell.value = value
            cell.font = font
            cell.alignment = align
            if border: cell.border = border
            if fill: cell.fill = fill
            return cell

        def merge_and_set(range_str, value, font=self.font_normal, align=self.align_center, border=self.border_thin, fill=None):
            ws.merge_cells(range_str)
            start_cell = range_str.split(':')[0]
            cell = set_cell(start_cell, value, font, align, None, fill)
            # Apply border to all cells in merged range to avoid missing lines
            for row in ws[range_str]:
                for c in row:
                    if border: c.border = border
            return cell

        # =========================================================================
        # HEADER SECTION
        # =========================================================================
        merge_and_set('A2:R2', "비파괴검사 결과서 및 작업/감독일보", font=self.font_title, border=None)
        
        # Left Info
        merge_and_set('A4:F4', "용 역 명 : 2026년 중앙지사 열수송관 비파괴검사용역", font=self.font_normal, align=self.align_left, border=None)
        date_str = data.get('date', '2026년   월   일')
        weather_str = data.get('weather', '')
        merge_and_set('A5:F5', f"검사일자 : {date_str}          날씨 : {weather_str}", font=self.font_normal, align=self.align_left, border=None)

        # Right Info (Signatures)
        
        set_cell('O4', '현장대리인', font=self.font_small, fill=self.fill_header)
        set_cell('P4', '감독', font=self.font_small, fill=self.fill_header)
        merge_and_set('Q4:R4', '확인', font=self.font_small, fill=self.fill_header)
        
        set_cell('O5', '') # Signature space
        set_cell('P5', '')
        merge_and_set('Q5:R5', '')
        ws.row_dimensions[5].height = 30 # Make signature box taller
        
        # Add signature image
        try:
            from openpyxl.drawing.image import Image
            sign_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'signs', '주진철.png')
            if os.path.exists(sign_path):
                img = Image(sign_path)
                # Resize image to fit the cell (approx width 70, height 38)
                img.width = 50
                img.height = 35
                
                # Use OneCellAnchor to center the image in O5
                from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
                from openpyxl.drawing.xdr import XDRPositiveSize2D
                from openpyxl.utils.units import pixels_to_EMU
                
                # O is column index 14, row 5 is index 4
                marker = AnchorMarker(col=14, colOff=pixels_to_EMU(12), row=4, rowOff=pixels_to_EMU(3))
                size = XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height))
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                
                ws.add_image(img)
        except Exception as e:
            print(f"Error adding signature: {e}")

        # =========================================================================
        # SECTION 1: 작업 물량 및 누계 현황 (A7:J22)
        # =========================================================================
        merge_and_set('A7:F7', "1. 작업 물량 및 누계 현황", font=self.font_bold, align=self.align_left, border=None)
        
        headers_qty = ['방법', '규격', '예상량', '전일 누계', '금일 작업', '총 누계', '공정률(%)', '불량', '불량률(%)', '비고']
        for i, header in enumerate(headers_qty):
            col_letter = get_column_letter(i+1)
            set_cell(f'{col_letter}8', header, font=self.font_bold, fill=self.fill_header, align=self.align_nowrap)

        qty_rows = [
            ('PAUT', '300A이상'), ('PAUT', '300A이상-야간'), ('PAUT', '250A'), ('PAUT', '200A'), ('PAUT', '200A-야간'), ('PAUT', '소계'),
            ('RT', '150A~100A'), ('RT', '150A~100A-야간'), ('RT', '80A이하'), ('RT', '80A이하-야간'), ('RT', '소계'),
            ('MT', '전체(주간)'), ('MT', '전체(야간)'),
            ('PT', '전체(주간)'), ('PT', '전체(야간)')
        ]
        
        row_idx = 9
        for method, spec in qty_rows:
            set_cell(f'A{row_idx}', method)
            set_cell(f'B{row_idx}', spec)
            
            # Fetch data from input dictionary
            row_data = data.get('qty_data', {}).get(f"{method}_{spec}", {})
            
            is_m_unit = method in ['PAUT', 'MT', 'PT']
            
            def set_qty_cell(coord, val_str, force_m_format=False):
                if not val_str:
                    set_cell(coord, '')
                    return
                try:
                    fval = float(str(val_str).replace(',', ''))
                    cell = ws[coord]
                    cell.value = fval
                    if force_m_format:
                        cell.number_format = '0.0000'
                    elif fval % 1 != 0:
                        cell.number_format = '0.0'
                    else:
                        cell.number_format = '0'
                    cell.font = self.font_normal
                    cell.alignment = self.align_center
                    cell.border = self.border_thin
                except ValueError:
                    set_cell(coord, val_str)

            set_qty_cell(f'C{row_idx}', row_data.get('예상량', ''), force_m_format=False)
            set_qty_cell(f'D{row_idx}', row_data.get('전일누계', ''), force_m_format=is_m_unit)
            set_qty_cell(f'E{row_idx}', row_data.get('금일작업', ''), force_m_format=is_m_unit)
            set_qty_cell(f'F{row_idx}', row_data.get('총누계', ''), force_m_format=is_m_unit)
            
            set_qty_cell(f'G{row_idx}', row_data.get('공정률', ''))
            set_qty_cell(f'H{row_idx}', row_data.get('불량', ''))
            set_qty_cell(f'I{row_idx}', row_data.get('불량률', ''))
            set_cell(f'J{row_idx}', row_data.get('비고', ''))
            
            row_idx += 1
            
        ws.merge_cells('A9:A14') # PAUT
        ws.merge_cells('A15:A19') # RT
        ws.merge_cells('A20:A21') # MT
        ws.merge_cells('A22:A23') # PT

        # =========================================================================
        # RIGHT SIDE SECTION: Equipment and Personnel (L7:P23)
        # =========================================================================
        # Equipment
        merge_and_set('L8:N8', '장비투입 현황(대)', font=self.font_bold, fill=self.fill_header)
        set_cell('L9', '장비명', font=self.font_bold, fill=self.fill_header)
        set_cell('M9', '금일', font=self.font_bold, fill=self.fill_header)
        set_cell('N9', '누계', font=self.font_bold, fill=self.fill_header)
        
        equip_rows = ['PAUT장비', 'PAUT프로브', 'PAUT스캐너', 'RT장비', 'MT장비']
        for i, eq in enumerate(equip_rows):
            r = 10 + i
            set_cell(f'L{r}', eq, align=self.align_nowrap)
            eq_data = data.get('equip_data', {}).get(eq, {})
            set_cell(f'M{r}', eq_data.get('금일', ''))
            set_cell(f'N{r}', eq_data.get('누계', ''))

        # Personnel
        merge_and_set('O8:R8', '금일 투입인원(명)', font=self.font_bold, fill=self.fill_header)
        set_cell('O9', '구분(관리/안전)', font=self.font_small, fill=self.fill_header, align=self.align_nowrap)
        set_cell('P9', '검사원', font=self.font_bold, fill=self.fill_header, align=self.align_nowrap)
        merge_and_set('Q9:R9', '안전담당', font=self.font_bold, fill=self.fill_header, align=self.align_nowrap)
        
        personnel = data.get('personnel_data', {})
        set_cell('O10', '인원', align=self.align_nowrap)
        set_cell('P10', personnel.get('검사원_인원', ''), align=self.align_nowrap)
        merge_and_set('Q10:R10', personnel.get('안전_인원', ''), align=self.align_nowrap)
        
        set_cell('O11', '현장대리인', align=self.align_nowrap)
        set_cell('P11', personnel.get('검사원_현장대리인', ''), align=self.align_nowrap)
        merge_and_set('Q11:R11', personnel.get('안전_현장대리인', ''), align=self.align_nowrap)
        
        set_cell('O12', '누계', align=self.align_nowrap)
        set_cell('P12', personnel.get('검사원_누계', ''), align=self.align_nowrap)
        merge_and_set('Q12:R12', personnel.get('안전_누계', ''), align=self.align_nowrap)
        
        # Merge empty blocks under personnel to match equipment height
        merge_and_set('O13:R14', '', border=self.border_thin)

        # Remarks (특이사항 및 작업계획)
        merge_and_set('L15:R15', '특이사항 및 작업계획', font=self.font_bold, fill=self.fill_header)
        merge_and_set('L16:R23', data.get('remarks', ''), align=Alignment(horizontal='left', vertical='top', wrap_text=True))

        # =========================================================================
        # SECTION 2: 비파괴검사결과서 (A26:Y...)
        # =========================================================================
        merge_and_set('A25:F25', "2. 비파괴검사결과서", font=self.font_bold, align=self.align_left, border=None)
        
        # Main Headers
        headers_ndt = [
            ('A26:A27', '순번'), ('B26:B27', '업체'), ('C26:C27', '검사방법'), ('D26:D27', '구간(Sec.No)'), ('E26:E27', '라인번호'),
            ('F26:F27', 'Joint No.'), ('G26:G27', '관경'), ('H26:H27', '두께'), ('I26:I27', '용접사'), 
            ('N26:N27', '결과'), ('O26:O27', '규격')
        ]
        for rng, text in headers_ndt:
            merge_and_set(rng, text, font=self.font_small, fill=self.fill_header, align=self.align_nowrap)
            
        merge_and_set('J26:M26', '구간정보(Start/Length)', font=self.font_small, fill=self.fill_header)
        set_cell('J27', '1', font=self.font_small, fill=self.fill_header)
        set_cell('K27', '2', font=self.font_small, fill=self.fill_header)
        set_cell('L27', '3', font=self.font_small, fill=self.fill_header)
        set_cell('M27', '4', font=self.font_small, fill=self.fill_header)
            
        merge_and_set('P26:Q26', 'RT매수', font=self.font_small, fill=self.fill_header)
        set_cell('P27', 'OR', font=self.font_small, fill=self.fill_header)
        set_cell('Q27', 'RE', font=self.font_small, fill=self.fill_header)
        
        merge_and_set('R26:R27', 'PAUT(m)', font=self.font_small, fill=self.fill_header)
        merge_and_set('S26:S27', 'MT(m)', font=self.font_small, fill=self.fill_header)
        merge_and_set('T26:T27', 'PT(m)', font=self.font_small, fill=self.fill_header)
        
        # Populate NDT Results
        ndt_results = data.get('ndt_results', [])
        start_row = 28
        total_rows = max(10, len(ndt_results) + 1)
        for i in range(total_rows): # minimum 10 empty rows for good look
            row_idx = start_row + i
            res = ndt_results[i] if i < len(ndt_results) else {}
            
            set_cell(f'A{row_idx}', i + 1)
            set_cell(f'B{row_idx}', res.get('업체', ''))
            set_cell(f'C{row_idx}', res.get('검사방법', ''))
            set_cell(f'D{row_idx}', res.get('구간', ''))
            set_cell(f'E{row_idx}', res.get('라인번호', ''))
            set_cell(f'F{row_idx}', res.get('Joint No.', ''))
            set_cell(f'G{row_idx}', res.get('관경', ''))
            set_cell(f'H{row_idx}', res.get('두께', ''))
            set_cell(f'I{row_idx}', res.get('용접사', ''))
            
            method = res.get('검사방법', '').upper().strip()
            
            sec_info_raw = res.get('구간정보', '').strip()
            
            # 검사방법이나 결과 등 데이터가 있는 실제 행인데 구간정보가 비었을 때 '관련지시 없음' 처리
            is_empty_row = not any(res.values())
            if not is_empty_row and not sec_info_raw:
                sec_info_raw = "관련지시 없음"
                
            sec_info = sec_info_raw.split(',')
            sec_info += [''] * (4 - len(sec_info)) # Pad to 4
            
            set_cell(f'J{row_idx}', sec_info[0])
            set_cell(f'K{row_idx}', sec_info[1] if method not in ['PAUT', 'PT', 'MT'] and sec_info_raw != "관련지시 없음" else '')
            set_cell(f'L{row_idx}', sec_info[2] if method not in ['PAUT', 'PT', 'MT'] and sec_info_raw != "관련지시 없음" else '')
            set_cell(f'M{row_idx}', sec_info[3] if method not in ['PAUT', 'PT', 'MT'] and sec_info_raw != "관련지시 없음" else '')
            
            if method in ['PAUT', 'PT', 'MT'] or sec_info_raw == "관련지시 없음":
                ws.merge_cells(f'J{row_idx}:M{row_idx}')
            
            spec_val = res.get('규격', '')
            shift_val = res.get('근무구분', '')
            if shift_val and shift_val != '주간':
                spec_val = f"{spec_val}({shift_val})" if spec_val else shift_val
                
            set_cell(f'N{row_idx}', res.get('결과', ''))
            set_cell(f'O{row_idx}', spec_val)
            set_cell(f'P{row_idx}', res.get('RT_OR', ''))
            set_cell(f'Q{row_idx}', res.get('RT_RE', ''))
            set_cell(f'R{row_idx}', res.get('PAUT', ''))
            set_cell(f'S{row_idx}', res.get('MT', ''))
            set_cell(f'T{row_idx}', res.get('PT', ''))
            
            if i == len(ndt_results):
                ws.merge_cells(f'B{row_idx}:T{row_idx}')
                ws[f'B{row_idx}'].value = "- 이 하 여 백 -"
                ws[f'B{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
                
            ws.row_dimensions[row_idx].height = 20

        # Print setup (Landscape, fit to one page width)
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = '26:27'
        
        # Header/Footer (Auto Page Numbering)
        ws.oddHeader.right.text = "Page &P of &N"
        ws.evenHeader.right.text = "Page &P of &N"
        
        # Margins
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.8
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0.3

        wb.save(output_path)
        return output_path

if __name__ == "__main__":
    # Test generation
    exporter = DailyWorkLogExporter()
    test_data = {
        'date': '2026년 08월 05일',
        'weather': '맑음',
        'qty_data': {
            '300A이상': {'예상량': 129, '전일누계': 100, '금일작업': 10, '총누계': 110, '공정률': '85.2%'}
        },
        'remarks': '안전 작업 지시 이행\nPAUT 장비 점검 완료',
        'ndt_results': [
            {'검사방법': 'PAUT', '구간': 'S-1', '라인번호': 'L-1', 'Joint No.': 'J-1', '결과': '합격'}
        ]
    }
    # exporter.generate_excel(test_data, "test_daily_report.xlsx")
