import json
import io
import os
import re
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


APP_TITLE = "지역난방 안전관리교육 자동 생성"
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config_district_heating_safety_training.json")
DEFAULT_ATTENDEES = (
    "주진철/서울검사/부장, 박광복/서울검사/부장, 권인경/서울검사/차장, "
    "우명광/서울검사/대리, 박상태/서울검사/주임, 장승대/서울검사/주임"
)
DEFAULT_CONTENT = """1. 작업 전 안전점검 및 작업구역 출입통제
 - 작업계획, 위험요인, 근제대책을 공유하고 작업자별 역할을 확인한다.
 - 안전표지, 차단선 및 감시자를 배치하고 무관계자의 출입을 통제한다.

2. 방사선투과검사(RT) 안전수칙
 - 작업 전 조사기, 가이드튜브, 원격조작기 및 측정기의 정상 작동을 확인한다.
 - 개인선량계를 착용하고 방사선 관리구역의 경계선량률을 확인한다.
 - 선원 인출·회수 상태를 확인한 후에만 통제를 해제한다.

3. 열수송관 현장 주요 재해 예방
 - 굴착부·개구부 추락, 남하, 협착 및 중량물 취급 위험을 사전에 점검한다.
 - 인양구와 슬링밸트의 손상·하중을 확인하고 인양물 아래에 진입하지 않는다.
 - 전기, 화재, 밀폐공간 등 현장 여건에 따른 보호구를 올바르게 착용한다.

4. 비상시 조치절차
 - 이상 발견 시 즉시 작업을 중지하고 안전한 장소로 대피한다.
 - 관리감독자에게 보고하고 현장 비상조치절차에 따라 대응한다.
 - 작업 종료 후 작업구역, 장비, 선량 및 교육기록을 확인한다."""

PROCESSES = ("RT", "PAUT", "PT", "MT")
COMMON_MONTHLY_TOPICS = [
    "신년 작업계획·위험성평가·작업중지권 확인",
    "겨울철 결빙·한랭·난방기구 화재 예방",
    "작업전 TBM과 근제대책의 현장 이행 확인",
    "해빙기 굴착부·개구부·지반침하 재해 예방",
    "중량물 취급·인양구·슬링밸트 안전점검",
    "여름철 온열질환 예방과 휴식·수분섭취 기준",
    "장마철 감전·침수·미끄럼 예방과 전기기구 점검",
    "폭염기 작업시간 조정·응급조치·집단휴식 기준",
    "추락·낙하·협착 위험구역 점검과 보호구 착용",
    "화재·누출·인명사고 비상대피 및 보고체계 훈련",
    "장비·계측기 정기점검과 불량장비 사용금지",
    "연말 사고사례 및 아차사고 공유·재발방지 대책",
]
UNDERGROUND_MONTHLY_TOPICS = [
    "굴착면·흑막이·지보공의 손상, 변형, 토사 균열 점검",
    "동결·해빙 반복에 따른 지반 이완과 굴착면 붕괴 예방",
    "지하 출입용 사다리·계단·가설통로의 고정·미끄럼 점검",
    "개구부 안전난간·추락방호와 상부 낙하물 통제",
    "굴착 상부 토사·자재·장비 적치위치와 추가하중 관리",
    "환기가 불량한 지하부의 산소·유해가스 사전측정과 환기",
    "우수·지하수·배관수 유입에 대비한 배수경로와 긴급대피",
    "굴착기·인양장비 가동 구역 접근 금지 및 상부 신호수 배치 확인",
    "협소한 배관 주변의 불안정한 작업자세·협착·탈출로 확보",
    "열수송관 누수·잔류압력·고온수 노출 시 작업중지와 대피",
    "지하 전기기구·연장선의 방수·접지·누전차단기 점검",
    "단독작업 금지, 상부 감시자, 비상연락·구조·현장보존 절차",
]
PROCESS_MONTHLY_TOPICS = {
    "RT": [
        "조사기·선원·원격조작기 사전점검", "방사선 관리구역·경고표지·출입통제",
        "개인선량계·서베이미터 착용과 작동확인", "선원 인출·회수 확인과 이상 시 응급조치",
        "가이드튜브·커넥터 손상과 걸림 예방", "감시자 배치와 작업자 간 신호체계",
        "비상용 선원회수도구·차폐체 비치와 사용법", "야간 RT 조명·통신·단독작업 금지",
        "경계선량률 측정·기록과 관리구역 해제 조건", "방사선 비상사고 대피·보고·현장보존",
        "방사선작업종사자 교육·검진·선량기록 관리", "RT 작업 아차사고와 선원분실 예방",
    ],
    "PAUT": [
        "탐촉자·웨지·케이블 사전점검", "스캐너 이동부 협착과 케이블 걸림 예방",
        "표면 전처리 시 그라인더·비산물 안전", "고소·협소부 PAUT 검사 자세와 추락 예방",
        "검사장비 전원·배선·누전차단기 점검", "커플런트 취급과 피부·바닥 미끄럼 예방",
        "열수송관 고온표면·잔류열 확인", "장시간 반복주사 작업의 근골격계 부담 예방",
        "검사구역 출입통제와 작업자 간 신호", "장비 낙하·전도 및 이동 중 충격 예방",
        "캘리브레이션 블록 취급과 장비 이상 조치", "PAUT 검사 아차사고와 장비손상 예방",
    ],
    "PT": [
        "침투액·세척액·현상액 SDS 및 경고표지", "화학물질 보호장갑·보안경 착용과 피부접촉 예방",
        "에어로졸 환기·화기관리·흡입 예방", "표면전처리 시 비산물·소음·진동 관리",
        "침투제 누출·유출과 바닥 미끄럼 예방", "밀폐·협소공간 PT 시 환기와 가스농도 확인",
        "고온 용접부 냉각 확인과 화상 예방", "폐기물·오염걸레 분리배출과 화재 예방",
        "작업장 조명·작업자세·접근통로 확보", "화학물질 눈·피부 노출 응급세척과 보고",
        "용기 라벨·유통기한·보관장소 점검", "PT 작업 아차사고와 화학물질 오용 예방",
    ],
    "MT": [
        "자화기·프로드·케이블 외관과 절연 점검", "전기식 MT 감전·단락·누전차단기 안전",
        "자분·현탁액 취급과 보안경·보호장갑 착용", "자화기 자력·협착·낙하 위험 예방",
        "표면전처리 그라인더 반발·비산물 안전", "습윤 현장의 전기장비·연장선 사용기준",
        "검사액 누출과 바닥 미끄럼·오염 예방", "고소·협소부 MT 작업자세와 추락 예방",
        "암실·자외선등 사용 시 눈·피부 보호", "자화전류 차단·장비 이상 시 응급조치",
        "탈자·장비정리·컨테이너 보관 안전", "MT 작업 아차사고와 감전·화재 예방",
    ],
}

PAUT_MONTHLY_CORE_SAFETY = [
    " - PAUT 본체·탐촉자·웨지·케이블의 외관과 정상 작동을 확인한다.\n"
    " - 기준시험편으로 웨지지연·감도·TCG를 교정한 후 검사를 시작한다.\n",
    " - 스캐너의 바퀴·체인·스프링 암에 손가락이 끼이지 않도록 한다.\n"
    " - 탐촉자·웨지 교체 전 전원과 커플런트 공급을 차단한다.\n",
    " - 표면전처리 시 그라인더 반발·비산물·소음을 예방하고 보안경·귀마개를 착용한다.\n"
    " - 연삭작업 전 배관표면의 잔류열과 주변 가연물을 확인한다.\n",
    " - 열수송관의 차단·감압·배수·잔류압력과 표면온도를 확인한다.\n"
    " - 고온표면에는 적합한 웨지·커플런트를 사용하고 화상을 예방한다.\n",
    " - 스캐너·탐촉자의 낙하방지를 실시하고 장비를 들고 사다리를 오르내리지 않는다.\n"
    " - 인양로프·용기를 사용하여 검사장비를 안전하게 반입·반출한다.\n",
    " - 장시간 웅크림·비틀림 자세를 피하고 작업발판과 교대작업을 활용한다.\n"
    " - 반복스캔 중 손목·허리의 근골격계 부담을 줄이기 위해 정기적으로 스트레칭한다.\n",
    " - 전동펌프·PAUT 장비·연장선의 접지, 방수, 누전차단기 작동을 확인한다.\n"
    " - 젖은 손으로 커넥터를 조작하지 않고 전원부에 커플런트가 닿지 않게 한다.\n",
    " - 누출된 커플런트를 즉시 제거하고 미끄럼방지 안전화를 착용한다.\n"
    " - 케이블·호스를 승강통로 밖으로 정리하여 걸림·넘어짐을 예방한다.\n",
    " - 인코더 원점·이동거리·스캔방향을 확인하고 누락구간이 없도록 한다.\n"
    " - 검사 중 케이블 풀림·노이즈·커플링 불량 발생 시 검사를 중지한다.\n",
    " - 배관 곡률과 용접부 형상에 맞게 스캐너를 고정하고 이탈·낙하를 예방한다.\n"
    " - 상부 감시자와 신호방법을 정하고 장비·자재 낙하구역에 진입하지 않는다.\n",
    " - 환기가 불량한 지하부는 진입 전·재진입 전에 산소·유해가스를 확인한다.\n"
    " - 상부 감시자와 무전기 연락을 유지하고 단독작업을 금지한다.\n",
    " - 데이터 저장 전 스캔범위·감도·TCG·인코더 기록과 누락구간을 재확인한다.\n"
    " - 검사 종료 후 탐촉자·웨지·케이블·커플런트를 회수하고 장비 이상을 기록한다.\n",
]


class SafetyTrainingApp:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("780x760")
        self.root.minsize(720, 640)
        self.photo_paths = []
        self.config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), CONFIG_FILE)
        self._build_ui()
        self._load_config()
        self._refresh_monthly_content()

    def _build_ui(self):
        outer = ttk.Frame(self.root, padding=14)
        outer.pack(fill="both", expand=True)

        ttk.Label(outer, text=APP_TITLE, font=("Malgun Gothic", 17, "bold")).pack(pady=(0, 12))
        form = ttk.LabelFrame(outer, text="교육 기본정보", padding=10)
        form.pack(fill="x")

        self.date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d 09:00"))
        self.location_var = tk.StringVar(value="한국지역난방공사 중앙지사 현장")
        self.type_var = tk.StringVar(value="작업 전 안전교육")
        self.duration_var = tk.StringVar(value="1시간")
        self.instructor_var = tk.StringVar()
        self.attendees_var = tk.StringVar(value=DEFAULT_ATTENDEES)
        self.year_var = tk.StringVar(value=str(datetime.now().year))
        self.month_var = tk.StringVar(value=f"{datetime.now().month:02d}")
        self.process_vars = {process: tk.BooleanVar(value=True) for process in PROCESSES}

        rows = [
            ("교육일시", self.date_var), ("교육장소", self.location_var),
            ("교육구분", self.type_var), ("교육시간", self.duration_var),
            ("강사", self.instructor_var), ("참석자", self.attendees_var),
        ]
        for row, (label, variable) in enumerate(rows):
            ttk.Label(form, text=label, width=12).grid(row=row, column=0, sticky="e", padx=5, pady=4)
            if label == "교육구분":
                widget = ttk.Combobox(form, textvariable=variable, state="readonly", values=(
                    "작업 전 안전교육", "정기 안전교육", "특별 안전교육", "비상대응 교육"
                ))
            else:
                widget = ttk.Entry(form, textvariable=variable)
            widget.grid(row=row, column=1, sticky="ew", padx=5, pady=4)
        form.columnconfigure(1, weight=1)
        ttk.Label(form, text="참석자는 '이름/소속/직책'을 쉼표로 구분", foreground="#555555").grid(
            row=6, column=1, sticky="w", padx=5
        )

        monthly = ttk.LabelFrame(outer, text="월별·공정별 교육 설정", padding=8)
        monthly.pack(fill="x", pady=(10, 0))
        ttk.Label(monthly, text="대상 연월").pack(side="left", padx=(3, 5))
        ttk.Spinbox(monthly, from_=2020, to=2100, textvariable=self.year_var, width=6).pack(side="left")
        ttk.Label(monthly, text="년").pack(side="left", padx=(2, 5))
        ttk.Spinbox(monthly, from_=1, to=12, format="%02.0f", textvariable=self.month_var, width=4).pack(side="left")
        ttk.Label(monthly, text="월").pack(side="left", padx=(2, 12))
        for process in PROCESSES:
            ttk.Checkbutton(monthly, text=process, variable=self.process_vars[process]).pack(side="left", padx=4)
        ttk.Button(monthly, text="월별 교육내용 작성", command=self._refresh_monthly_content).pack(side="right")

        buttons = ttk.Frame(outer)
        buttons.pack(fill="x", pady=(10, 0))
        ttk.Button(buttons, text="교육사진 선택", command=self._select_photos).pack(side="left")
        self.photo_label = ttk.Label(buttons, text="선택된 사진 없음")
        self.photo_label.pack(side="left", padx=10)
        ttk.Button(
            buttons,
            text="Excel 교육일지 생성",
            command=self.generate_excel,
        ).pack(side="right")

        content_frame = ttk.LabelFrame(outer, text="교육내용", padding=8)
        content_frame.pack(fill="both", expand=True, pady=10)
        self.content_text = tk.Text(content_frame, height=12, wrap="word", font=("Malgun Gothic", 10))
        self.content_text.pack(fill="both", expand=True)
        self.content_text.insert("1.0", DEFAULT_CONTENT)

        action_frame = ttk.LabelFrame(outer, text="교육결과 및 조치사항", padding=8)
        action_frame.pack(fill="x")
        self.action_text = tk.Text(action_frame, height=3, wrap="word", font=("Malgun Gothic", 10))
        self.action_text.pack(fill="x")
        self.action_text.insert("1.0", "교육내용 전파 및 참석자 이해도 확인. 특이사항 없음.")
        self._refresh_monthly_content()

    def _selected_processes(self):
        return [process for process in PROCESSES if self.process_vars[process].get()]

    def _monthly_content_for(self, process, month):
        common_topic = COMMON_MONTHLY_TOPICS[month - 1]
        process_topic = PROCESS_MONTHLY_TOPICS[process][month - 1]
        underground_topic = UNDERGROUND_MONTHLY_TOPICS[month - 1]
        process_core_safety = PAUT_MONTHLY_CORE_SAFETY[month - 1] if process == "PAUT" else ""
        return (
            f"1. {month}월 공통 안전교육\n"
            f" - {common_topic}\n"
            " - 작업 전 TBM에서 작업순서, 위험요인, 안전대책과 작업중지 기준을 확인한다.\n"
            " - 작업구역 출입통제, 보호구 착용, 비상연락체계를 공유한다.\n\n"
            f"2. {process} 공정 중점교육\n"
            f" - {process_topic}\n"
            f"{process_core_safety}"
            " - 사용 장비와 계측기의 점검상태를 확인하고, 이상 발견 시 즉시 사용을 중지하고 장비 재점검을 실시한다.\n"
            " - 작업 중 이상상태 발견 시 작업을 중지하고 관리감독자에게 보고한다.\n\n"
            "3. 지하 1.5~3m 배관 용접부 검사 공통위험\n"
            f" - {underground_topic}\n"
            " - 안전한 승강통로(사다리, 가설계단 등)가 확보된 상태에서만 진입한다.\n"
            " - 무거운 검사장비는 무리하게 들고 승강하지 않으며, 달줄이나 달포대 등을 이용해 안전하게 반입한다.\n"
            " - 붕괴 조짐, 침수, 유해가스 경보 발생 또는 중장비 접근 시 즉시 밖으로 대피한다.\n\n"
            "4. 교육확인\n"
            " - 참석자에게 중점사항을 질문하여 이해도를 확인하고 미흡사항은 재교육한다.\n"
            " - 작업 종료 후 인원, 장비 및 사용 자재의 철수를 확인하고 개구부 재차단을 확인한다."
        )

    def _refresh_monthly_content(self):
        try:
            year = int(self.year_var.get())
            month = int(self.month_var.get())
            if not 1 <= month <= 12:
                raise ValueError
        except ValueError:
            messagebox.showwarning("연월 확인", "올바른 대상 연도와 1~12월을 입력해 주세요.")
            return
        selected = self._selected_processes()
        if not selected:
            messagebox.showwarning("공정 확인", "RT, PAUT, PT, MT 중 하나 이상을 선택해 주세요.")
            return
        sections = [f"[{year}년 {month:02d}월 지역난방 안전관리교육]"]
        for process in selected:
            sections.append(f"\n■ {process} 공정\n{self._monthly_content_for(process, month)}")
        self.content_text.delete("1.0", "end")
        self.content_text.insert("1.0", "\n".join(sections))

    def _select_photos(self):
        paths = filedialog.askopenfilenames(
            title="교육사진 선택", filetypes=[("이미지", "*.png;*.jpg;*.jpeg;*.bmp")]
        )
        if paths:
            self.photo_paths = list(paths)
            self.photo_label.config(text=f"{len(paths)}개 사진 선택됨")

    def _load_config(self):
        if not os.path.exists(self.config_path):
            return
        try:
            with open(self.config_path, "r", encoding="utf-8") as file:
                data = json.load(file)
            self.location_var.set(data.get("location", self.location_var.get()))
            self.instructor_var.set(data.get("instructor", ""))
            for process in PROCESSES:
                self.process_vars[process].set(data.get("processes", {}).get(process, True))
        except (OSError, ValueError):
            pass

    def _save_config(self):
        try:
            with open(self.config_path, "w", encoding="utf-8") as file:
                json.dump({
                    "location": self.location_var.get(),
                    "instructor": self.instructor_var.get(),
                    "processes": {process: variable.get() for process, variable in self.process_vars.items()},
                }, file, ensure_ascii=False, indent=2)
        except OSError:
            pass

    @staticmethod
    def _parse_attendees(value):
        result = []
        for item in re.split(r"[,\n]", value):
            item = item.strip()
            if not item:
                continue
            if "/" in item and "(" not in item:
                name, role = item.split("/", 1)
                result.append((name.strip(), role.strip()))
                continue
            match = re.match(r"^(.*?)(?:\((.*?)\))?$", item)
            result.append((match.group(1).strip(), (match.group(2) or "").strip()))
        return result

    @staticmethod
    def _style_range(ws, cell_range, border, font, alignment=None, fill=None):
        for row in ws[cell_range]:
            for cell in row:
                cell.border = border
                cell.font = font
                if alignment:
                    cell.alignment = alignment
                if fill:
                    cell.fill = fill

    def _make_detail_sheet(self, wb, attendees, process, content, first_sheet=False):
        ws = wb.active if first_sheet else wb.create_sheet()
        ws.title = f"{process} 안전교육"
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1 if not self.photo_paths else 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = not self.photo_paths
        ws.page_margins.left = ws.page_margins.right = 0.35
        ws.page_margins.top = ws.page_margins.bottom = 0.45
        for col, width in zip("ABCDEF", [16, 18, 18, 16, 18, 18]):
            ws.column_dimensions[col].width = width

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        normal = Font(name="Malgun Gothic", size=10)
        bold = Font(name="Malgun Gothic", size=10, bold=True)
        title = Font(name="Malgun Gothic", size=19, bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
        gray = PatternFill("solid", fgColor="D9E2F3")
        light = PatternFill("solid", fgColor="EAF2F8")

        ws.merge_cells("A1:F2")
        ws["A1"] = "지역난방 안전관리교육 일지"
        ws["A1"].font = title
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 20

        info = [
            (3, "교육일시", self.date_var.get(), "교육구분", f"{process} / {self.type_var.get()}"),
            (4, "교육장소", self.location_var.get(), "교육시간", self.duration_var.get()),
            (5, "강사", self.instructor_var.get(), "참석인원", len(attendees)),
        ]
        for row, left_label, left_value, right_label, right_value in info:
            ws[f"A{row}"] = left_label
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            ws[f"B{row}"] = left_value
            ws[f"D{row}"] = right_label
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            ws[f"E{row}"] = right_value
            for label_cell in (ws[f"A{row}"], ws[f"D{row}"]):
                label_cell.fill = gray
                label_cell.font = bold
            self._style_range(ws, f"A{row}:F{row}", border, normal, center)
            ws.row_dimensions[row].height = 25

        ws.merge_cells("A7:F7")
        ws["A7"] = "1. 교육내용"
        ws["A7"].font = Font(name="Malgun Gothic", size=12, bold=True)
        ws["A7"].fill = light
        ws["A7"].alignment = Alignment(vertical="center")
        ws.merge_cells("A8:F8")
        ws["A8"] = content
        ws["A8"].font = normal
        ws["A8"].alignment = left_top
        self._style_range(ws, "A7:F8", border, normal)
        ws["A7"].font = Font(name="Malgun Gothic", size=12, bold=True)
        ws.row_dimensions[7].height = 27
        content_line_count = len(content.splitlines())
        ws.row_dimensions[8].height = max(330, min(480, content_line_count * 15))

        ws.merge_cells("A10:F10")
        ws["A10"] = "2. 교육결과 및 조치사항"
        ws["A10"].font = Font(name="Malgun Gothic", size=12, bold=True)
        ws["A10"].fill = light
        ws.merge_cells("A11:F11")
        ws["A11"] = self.action_text.get("1.0", "end").strip()
        ws["A11"].alignment = left_top
        self._style_range(ws, "A10:F11", border, normal)
        ws["A10"].font = Font(name="Malgun Gothic", size=12, bold=True)
        ws.row_dimensions[10].height = 27
        ws.row_dimensions[11].height = 55

        ws.merge_cells("A13:F13")
        ws["A13"] = "3. 교육 참석자 확인"
        ws["A13"].font = Font(name="Malgun Gothic", size=12, bold=True)
        ws["A13"].fill = light
        headers = ["성명", "소속/직책", "서명", "성명", "소속/직책", "서명"]
        for col, value in enumerate(headers, 1):
            cell = ws.cell(14, col, value)
            cell.font, cell.fill, cell.alignment, cell.border = bold, gray, center, border
        rows_needed = max(4, (len(attendees) + 1) // 2)
        for index in range(rows_needed):
            row = 15 + index
            for col in range(1, 7):
                ws.cell(row, col).border = border
                ws.cell(row, col).font = normal
                ws.cell(row, col).alignment = center
            for side, attendee_index in ((0, index * 2), (3, index * 2 + 1)):
                if attendee_index < len(attendees):
                    name, role = attendees[attendee_index]
                    ws.cell(row, 1 + side, name)
                    ws.cell(row, 2 + side, role)
                    self._insert_signature(ws, name, row, 3 + side)
            ws.row_dimensions[row].height = 28
        end_row = 14 + rows_needed
        self._style_range(ws, f"A13:F{end_row}", border, normal)
        ws["A13"].font = Font(name="Malgun Gothic", size=12, bold=True)

        if self.photo_paths:
            photo_start = end_row + 2
            ws.merge_cells(start_row=photo_start, start_column=1, end_row=photo_start, end_column=6)
            ws.cell(photo_start, 1, "4. 교육사진")
            ws.cell(photo_start, 1).font = Font(name="Malgun Gothic", size=12, bold=True)
            ws.cell(photo_start, 1).fill = light
            photo_row = photo_start + 1
            ws.merge_cells(start_row=photo_row, start_column=1, end_row=photo_row, end_column=6)
            ws.row_dimensions[photo_row].height = 260
            self._style_range(ws, f"A{photo_start}:F{photo_row}", border, normal)
            self._insert_photo(ws, self.photo_paths[0], f"B{photo_row}")
            end_row = photo_row

        self._style_range(ws, f"A1:F{end_row}", border, normal)
        ws["A1"].font = title
        ws["A1"].alignment = center
        ws["A1"].border = border
        ws.print_area = f"A1:F{end_row}"
        ws.oddFooter.center.text = "Page &P of &N"
        return ws

    def _make_summary_sheet(self, wb, process_contents):
        ws = wb.create_sheet("교육현황")
        headers = ["일자", "구분", "내용", "시간", "강사", "장소"]
        ws.append(headers)
        for process, content in process_contents.items():
            lines = content.strip().splitlines()
            summary_text = " / ".join(line.strip(" -") for line in lines if line.strip())
            ws.append([self.date_var.get(), f"{process} {self.type_var.get()}", summary_text,
                       self.duration_var.get(), self.instructor_var.get(), self.location_var.get()])
        widths = [20, 20, 70, 12, 16, 35]
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
            for row in range(1, len(process_contents) + 2):
                cell = ws.cell(row, col)
                cell.border = border
                cell.font = Font(name="Malgun Gothic", size=10, bold=(row == 1))
                cell.alignment = Alignment(horizontal="center" if row == 1 else "left", vertical="center", wrap_text=True)
                if row == 1:
                    cell.fill = PatternFill("solid", fgColor="D9E2F3")
        ws.row_dimensions[1].height = 26
        for row in range(2, len(process_contents) + 2):
            ws.row_dimensions[row].height = 65
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:F{len(process_contents) + 1}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_area = f"A1:F{len(process_contents) + 1}"

    def _make_safety_checklist_sheet(self, wb, selected_processes):
        ws = wb.create_sheet("작업전 안전점검표")
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.page_margins.left = ws.page_margins.right = 0.3
        ws.page_margins.top = ws.page_margins.bottom = 0.4

        widths = [15, 56, 13, 25, 16]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        normal = Font(name="Malgun Gothic", size=9)
        bold = Font(name="Malgun Gothic", size=9, bold=True)
        title_font = Font(name="Malgun Gothic", size=17, bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        gray = PatternFill("solid", fgColor="D9E2F3")

        ws.merge_cells("A1:E2")
        ws["A1"] = "지하 1.5~3m 배관 용접부 검사 작업 전 안전점검표"
        ws["A1"].font = title_font
        ws["A1"].alignment = center

        info = [
            (3, "작업일시", self.date_var.get(), "검사공정", ", ".join(selected_processes)),
            (4, "작업장소", self.location_var.get(), "작업깊이", "1.5~3m"),
            (5, "작업책임자", self.instructor_var.get(), "상부감시자", ""),
        ]
        for row, label1, value1, label2, value2 in info:
            ws.cell(row, 1, label1)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            ws.cell(row, 2, value1)
            ws.cell(row, 4, label2)
            ws.cell(row, 5, value2)
            for col in range(1, 6):
                cell = ws.cell(row, col)
                cell.border = border
                cell.font = bold if col in (1, 4) else normal
                cell.alignment = center
                if col in (1, 4):
                    cell.fill = gray

        headers = ["구분", "점검항목", "점검결과", "미흡사항·조치내용", "확인자"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(7, col, header)
            cell.font, cell.fill, cell.alignment, cell.border = bold, gray, center, border

        checklist = [
            ("작업허가", "작업허가서·당일 위험성평가·TBM을 실시했는가?"),
            ("인원", "작업책임자·상부감시자·작업자의 역할과 연락방법을 확인했는가?"),
            ("매설물", "전기·가스·통신·열수송관 등 지하매설물 위치를 확인했는가?"),
            ("배관격리", "열수송관 차단·감압·배수·잔류열 및 오조작 방지표지를 확인했는가?"),
            ("굴착부", "굴착면·흑막이·지보공에 균열, 변형, 부식, 토사탈락이 없는가?"),
            ("상부하중", "굴착 가장자리 주변의 토사·자재·장비로 인한 추가하중을 통제했는가?"),
            ("추락·낙하", "개구부 안전난간·추락방호·낙하물 방지·출입통제가 적정한가?"),
            ("승강통로", "사다리·계단을 견고하게 고정하고 미끄럼·파손 여부를 점검했는가?"),
            ("공기상태", "환기가 불량하거나 밀폐공간에 해당하는지 평가하고 산소·유해가스를 측정했는가?"),
            ("환기·구조", "환기장치·무전기·구조장비·비상탈출로를 확보했는가?"),
            ("유수·침수", "빗물·지하수·배관수 유입가능성과 배수·긴급대피 방법을 확인했는가?"),
            ("장비접촉", "굴착기·인양장비 작업반경을 분리하고 신호수를 배치했는가?"),
            ("장비반입", "검사장비를 사다리로 운반하지 않고 적합한 인양수단을 준비했는가?"),
            ("전기", "전기기구·연장선의 손상, 접지, 방수 및 누전차단기 작동을 확인했는가?"),
            ("공정장비", "선택한 RT·PAUT·PT·MT 장비와 계측기의 사전점검을 완료했는가?"),
            ("보호구", "안전모·안전화·반광조끼 및 공정별 추가 보호구를 착용했는가?"),
            ("조명·통로", "작업조명이 충분하고 피난·반출통로에 장애물이 없는가?"),
            ("기상·환경", "강우·강풍·폭염·한파·지반진동 등 작업중지 조건을 확인했는가?"),
            ("비상대응", "작업중지·대피·구조·보고체계와 119 안내위치를 공유했는가?"),
            ("종료확인", "작업 종료 후 인원·장비·선원·화학물질 회수와 개구부 재차단을 확인했는가?"),
        ]
        start_row = 8
        for index, (category, item) in enumerate(checklist, start_row):
            ws.cell(index, 1, category)
            ws.cell(index, 2, item)
            for col in range(1, 6):
                cell = ws.cell(index, col)
                cell.border = border
                cell.font = normal
                cell.alignment = center if col != 2 else left
            ws.row_dimensions[index].height = 27

        validation = DataValidation(type="list", formula1='"양호,불량,해당없음"', allow_blank=True)
        validation.promptTitle = "점검결과"
        validation.prompt = "양호, 불량, 해당없음 중 선택하세요."
        ws.add_data_validation(validation)
        validation.add(f"C{start_row}:C{start_row + len(checklist) - 1}")

        signature_row = start_row + len(checklist) + 1
        ws.cell(signature_row, 1, "최종확인")
        ws.merge_cells(start_row=signature_row, start_column=2, end_row=signature_row, end_column=3)
        ws.cell(signature_row, 2, "□ 작업가능    □ 보완 후 작업    □ 작업금지")
        ws.cell(signature_row, 4, "관리감독자")
        for col in range(1, 6):
            cell = ws.cell(signature_row, col)
            cell.border, cell.font, cell.alignment = border, bold, center
            if col in (1, 4):
                cell.fill = gray
        ws.row_dimensions[signature_row].height = 32
        ws.print_area = f"A1:E{signature_row}"
        ws.oddFooter.center.text = "Page &P of &N"

    @staticmethod
    def _insert_photo(ws, path, anchor):
        if not os.path.exists(path):
            return
        with PILImage.open(path) as image:
            width, height = image.size
        scale = min(560 / width, 300 / height, 1.0)
        picture = XLImage(path)
        picture.width = int(width * scale)
        picture.height = int(height * scale)
        ws.add_image(picture, anchor)

    @staticmethod
    def _insert_signature(ws, name, row, column):
        if not name:
            return False
        signs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "signs")
        candidates = (
            f"{name}_padded.png", f"{name}.png", f"{name}.jpg", f"{name}.jpeg"
        )
        path = next(
            (os.path.join(signs_dir, filename) for filename in candidates
             if os.path.exists(os.path.join(signs_dir, filename))),
            None,
        )
        if not path:
            return False

        # 서명 셀(C/F열: 너비 18, 행 높이 28pt)과 비슷한 투명 캔버스를
        # 만들어 서명이 셀 정중앙에 보이도록 한다.
        canvas_width, canvas_height = 126, 36
        with PILImage.open(path) as source:
            source = source.convert("RGBA")
            ratio = min(66 / source.width, 28 / source.height, 1.0)
            resized = source.resize(
                (max(1, int(source.width * ratio)), max(1, int(source.height * ratio))),
                PILImage.Resampling.LANCZOS,
            )
            canvas = PILImage.new("RGBA", (canvas_width, canvas_height), (255, 255, 255, 0))
            x = (canvas_width - resized.width) // 2
            y = (canvas_height - resized.height) // 2
            canvas.paste(resized, (x, y), resized)

        stream = io.BytesIO()
        canvas.save(stream, format="PNG")
        stream.seek(0)
        signature = XLImage(stream)
        signature.width = canvas_width
        signature.height = canvas_height
        ws.add_image(signature, f"{get_column_letter(column)}{row}")
        return True

    def generate_excel(self):
        if not self.date_var.get().strip() or not self.location_var.get().strip():
            messagebox.showwarning("입력 확인", "교육일시와 교육장소를 입력해 주세요.")
            return
        selected = self._selected_processes()
        if not selected:
            messagebox.showwarning("공정 확인", "RT, PAUT, PT, MT 중 하나 이상을 선택해 주세요.")
            return
        try:
            year = int(self.year_var.get())
            month = int(self.month_var.get())
            if not 1 <= month <= 12:
                raise ValueError
        except ValueError:
            messagebox.showwarning("연월 확인", "올바른 대상 연도와 1~12월을 입력해 주세요.")
            return
        attendees = self._parse_attendees(self.attendees_var.get())
        process_contents = {
            process: self._monthly_content_for(process, month) for process in selected
        }
        wb = openpyxl.Workbook()
        for index, (process, content) in enumerate(process_contents.items()):
            self._make_detail_sheet(wb, attendees, process, content, first_sheet=(index == 0))
        self._make_safety_checklist_sheet(wb, selected)
        self._make_summary_sheet(wb, process_contents)
        default_name = f"지역난방_안전관리교육_{year}{month:02d}.xlsx"
        path = filedialog.asksaveasfilename(
            title="지역난방 안전관리교육 저장", defaultextension=".xlsx",
            initialfile=default_name, filetypes=[("Excel 파일", "*.xlsx")]
        )
        if not path:
            return
        try:
            wb.save(path)
            self._save_config()
            messagebox.showinfo("완료", f"교육일지를 생성했습니다.\n{path}")
        except OSError as error:
            messagebox.showerror("저장 오류", str(error))


if __name__ == "__main__":
    root = tk.Tk()
    SafetyTrainingApp(root)
    root.mainloop()
