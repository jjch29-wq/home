import openpyxl
import json
import os

def find_paut_section(ws):
    """
    '1.2.1' 또는 '위상배열초음파' 섹션 제목을 찾고,
    그 아래에 있는 헤더 행과 데이터 시작 행, 열 매핑을 반환합니다.
    """
    title_rows = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value.strip()
                # 섹션 제목 행 찾기: '1.2.1'로 시작하거나 위상배열초음파 포함
                if '1.2.1' in v and ('위상' in v or 'PAUT' in v or '초음파' in v):
                    title_rows.append(cell.row)

    if not title_rows:
        return None, None, {}

    # 가장 마지막 1.2.1 제목을 사용 (실제 검사현황 표가 아래에 있을 가능성)
    for title_row in reversed(title_rows):
        # 제목 행 아래 20행 안에서 'Section'과 'Line No.' 를 모두 가진 헤더행 탐색
        for hr in range(title_row + 1, title_row + 20):
            row_vals = {}
            for col in range(1, 30):
                c = ws.cell(row=hr, column=col)
                if c.value and isinstance(c.value, str):
                    cv = c.value.strip().replace('\n', ' ')
                    if cv:
                        row_vals[col] = cv

            has_section = any('Section' in v for v in row_vals.values())
            has_lineno = any('Line No' in v for v in row_vals.values())

            if has_section and has_lineno:
                # 헤더 행 찾음! 컬럼 매핑 생성
                col_map = {}

                for col, val in row_vals.items():
                    val_lower = val.lower()
                    if val in ('업체', '회사') or '업체' in val:
                        col_map['company'] = col
                    elif val in ('순번', '번호', 'No', 'No.') or val.startswith('순번') or val.startswith('번호'):
                        col_map['no'] = col
                    elif 'Section' in val or '구간' in val:
                        col_map['section'] = col
                    elif 'Line No' in val or '라인' in val:
                        col_map['line_no'] = col
                    elif '관경' in val or '구경' in val or '직경' in val:
                        col_map['pipe_size'] = col
                    elif '용접개소' in val or 'Joint' in val or '조인트' in val:
                        col_map['joint'] = col
                    elif '규격' in val or '주야' in val:
                        col_map['shift'] = col
                    elif '단위' in val:
                        col_map['unit'] = col
                    elif 'PAUT' in val and ('길이' in val or '성적' in val or '검사' in val):
                        col_map['paut_start'] = col
                    elif '결과' in val:
                        col_map['result'] = col
                    elif '비고' in val:
                        col_map['note'] = col

                # 서브헤더 행 (ORI/RE/TOTAL 찾기)
                for sub_hr in range(hr + 1, hr + 3):
                    for col in range(1, 30):
                        c = ws.cell(row=sub_hr, column=col)
                        if c.value and isinstance(c.value, str):
                            cv = c.value.strip().replace("'", "")
                            if cv == 'ORI':
                                col_map['ori'] = col
                            elif cv == 'RE':
                                col_map['re'] = col
                            elif 'TOTAL' in cv:
                                col_map['total'] = col

                data_start = hr + 2  # 헤더 + 서브헤더 다음 행
                return hr, data_start, col_map

    return None, None, {}


def write_paut_data(ws, paut_records, header_row, data_start, col_map):
    """
    paut_records를 ws에 header_row 기반으로 기입합니다.
    병합 셀을 피하기 위해 항상 각 병합 범위의 첫 번째(앵커) 셀에만 값을 씁니다.
    """
    def safe_write(row, col, value):
        """병합 셀 충돌 없이 안전하게 값 쓰기"""
        for merge in ws.merged_cells.ranges:
            if merge.min_row <= row <= merge.max_row and merge.min_col <= col <= merge.max_col:
                # 앵커 셀(병합 범위의 왼쪽 위)에만 쓰기
                ws.cell(row=merge.min_row, column=merge.min_col).value = value
                return
        ws.cell(row=row, column=col).value = value

    # 기존 데이터 행 수 파악 (빈 행이 나올 때까지)
    existing_rows = 0
    for r in range(data_start, data_start + 200):
        row_has_data = False
        for c in range(1, 25):
            cell = ws.cell(row=r, column=c)
            if cell.value:
                row_has_data = True
                break
        if row_has_data:
            existing_rows += 1
        else:
            break

    # 새로운 행이 기존보다 많으면 행 삽입
    if len(paut_records) > max(existing_rows, 1):
        insert_count = len(paut_records) - max(existing_rows, 1)
        ws.insert_rows(data_start + 1, amount=insert_count)

    # 데이터 쓰기
    current_row = data_start
    for i, (key, data) in enumerate(paut_records.items()):
        업체, 구간, 라인번호, 관경, joint, 용접사 = key

        if 'company' in col_map:
            safe_write(current_row, col_map['company'], 업체)
        if 'no' in col_map:
            safe_write(current_row, col_map['no'], i + 1)
        if 'section' in col_map:
            safe_write(current_row, col_map['section'], 구간)
        if 'line_no' in col_map:
            safe_write(current_row, col_map['line_no'], 라인번호)
        if 'pipe_size' in col_map:
            safe_write(current_row, col_map['pipe_size'], 관경)
        if 'joint' in col_map:
            safe_write(current_row, col_map['joint'], joint)
        if 'shift' in col_map:
            safe_write(current_row, col_map['shift'], data.get('shift', '주간'))
        if 'unit' in col_map:
            safe_write(current_row, col_map['unit'], 'M')
        if 'result' in col_map:
            safe_write(current_row, col_map['result'], '합격')

        ori = data.get('ORI', 0.0)
        re_val = data.get('RE', 0.0)
        tot = ori + re_val

        if 'ori' in col_map and ori > 0:
            safe_write(current_row, col_map['ori'], round(ori, 4))
        if 're' in col_map and re_val > 0:
            safe_write(current_row, col_map['re'], round(re_val, 4))
        if 'total' in col_map and tot > 0:
            safe_write(current_row, col_map['total'], round(tot, 4))

        current_row += 1

    return current_row - data_start  # 작성된 행 수 반환


# ====================
# 테스트 실행
# ====================
if __name__ == '__main__':
    history_path = r'c:\Users\jjch2\Desktop\PMI\home\src\daily_work_history.json'
    template_path = r'C:\Users\jjch2\Desktop\템플릿_최종완성본_V74.xlsx'
    output_path = r'C:\Users\jjch2\Desktop\Test_HeaderBased.xlsx'
    target_month = '2026-08'

    with open(history_path, 'r', encoding='utf-8') as f:
        history = json.load(f)

    # PAUT 데이터 추출
    paut_raw = []
    for date_key, log_data in history.items():
        if date_key.startswith(target_month):
            for r in log_data.get('ndt_results', []):
                if str(r.get('검사방법', '')).strip().upper() == 'PAUT':
                    r['_date'] = date_key
                    paut_raw.append(r)

    paut_raw.sort(key=lambda x: x['_date'])

    groups = {}
    for r in paut_raw:
        key = (
            str(r.get('업체', '')),
            str(r.get('구간', '')),
            str(r.get('라인번호', '')),
            str(r.get('관경', '')),
            str(r.get('Joint No.', '')),
            str(r.get('용접사', ''))
        )
        paut_val = str(r.get('PAUT', '0')).strip()
        try:
            val = float(paut_val)
        except:
            val = 0.0
        if val == 0:
            continue

        shift = str(r.get('규격', '주간')).strip()
        if key not in groups:
            groups[key] = {'ORI': 0.0, 'RE': 0.0, 'shift': shift}
        if groups[key]['ORI'] == 0.0:
            groups[key]['ORI'] += val
        else:
            groups[key]['RE'] += val

    print(f'Found {len(groups)} PAUT groups')

    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]

    header_row, data_start, col_map = find_paut_section(ws)
    print(f'Header row: {header_row}, Data starts at: {data_start}')
    print(f'Column map: {col_map}')

    if header_row:
        written = write_paut_data(ws, groups, header_row, data_start, col_map)
        print(f'Wrote {written} rows of PAUT data')
        wb.save(output_path)
        print(f'Saved to {output_path}')
    else:
        print('PAUT table not found!')
