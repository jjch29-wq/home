### VERSION: BUDGET_SYNC_FIXED_FINAL_V8_1WON_BALANCED ###
import mimetypes
import os
# [OPTIMIZATION] Prevent slow Windows registry scan for mimetypes
if os.name == 'nt':
    mimetypes.init(files=[])

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sys
import subprocess
import os
import time
import datetime
import json
import ctypes
import re
import traceback
import sys
import subprocess
sys.path.append(os.path.abspath(os.path.dirname(__file__)))
from site_apps.central.src.utils.helpers import install_and_import, normalize_id
from site_apps.central.src.utils.helpers import NAN_PATTERN, DOT_ZERO_PATTERN, MARKER_PATTERN
# [FIX] Bypass slow pytz timezone loading on network drives (Google Drive)
import builtins
import inspect

_orig_open = builtins.open
class _PytzDummyFile:
    def close(self): pass

def _fast_open(file, mode='r', *args, **kwargs):
    if isinstance(file, str) and 'pytz' in file and 'zoneinfo' in file and mode == 'rb':
        try:
            frame = inspect.currentframe()
            if frame and frame.f_back and frame.f_back.f_code.co_name == 'open_resource':
                if frame.f_back.f_back and frame.f_back.f_back.f_code.co_name == 'resource_exists':
                    return _PytzDummyFile()
        except Exception:
            pass
    return _orig_open(file, mode, *args, **kwargs)

builtins.open = _fast_open

try:
    # Pre-import critical libraries
    pd = install_and_import('pandas')
    np = install_and_import('numpy')
    install_and_import('openpyxl')
    install_and_import('tkcalendar')
    install_and_import('xlsxwriter')
    install_and_import('pillow', 'PIL')
    from PIL import Image, ImageTk
finally:
    builtins.open = _orig_open

from site_apps.central.src.daily_work_report_manager import DailyWorkReportManager
from site_apps.central.src.ndt_billing_tab import NDTCalculatorTab
from site_apps.central.src.daily_work_log_tab import DailyWorkLogTab
import site_apps.central.src.daily_work_report_manager as daily_work_report_manager
print(f"DEBUG: daily_work_report_manager path: {daily_work_report_manager.__file__}")

# Pre-compiled regex for performance
from site_apps.central.src.views.components import *

# Auto-install additional dependencies
from tkcalendar import DateEntry, Calendar

# [FIX] Patch tkcalendar for stability in ko_KR locale / Python 3.14.
# 각 패치를 독립적으로 적용 - 한 패치 실패가 나머지를 무효화하지 않도록 분리.
import tkinter as _tk_ref

# 1) Calendar.__init__ : select_on_nav=False 기본값 + 헤더 너비 조정
try:
    _orig_cal_init = Calendar.__init__
    def _patched_cal_init(self, *args, **kwargs):
        if 'select_on_nav' not in kwargs:
            kwargs['select_on_nav'] = False
        _orig_cal_init(self, *args, **kwargs)
        def widen():
            if hasattr(self, '_header_month'):
                try: self._header_month.configure(width=25)
                except: pass
            if hasattr(self, '_header_year'):
                try: self._header_year.configure(width=10)
                except: pass
        self.after_idle(widen)
        self._widen_headers = widen
    Calendar.__init__ = _patched_cal_init
except Exception:
    pass

# 2) Calendar._display_calendar : 헤더 너비 재조정
try:
    _orig_cal_display = Calendar._display_calendar
    def _patched_cal_display(self, *args, **kwargs):
        _orig_cal_display(self, *args, **kwargs)
        if hasattr(self, '_widen_headers'):
            self.after_idle(self._widen_headers)
    Calendar._display_calendar = _patched_cal_display
except Exception:
    pass

# 3) DateEntry._show_calendar : tkcalendar < 1.6 버전 호환 (1.6.1에서는 없음 - 무시)
try:
    if hasattr(DateEntry, '_show_calendar'):
        _orig_de_show_cal = DateEntry._show_calendar
        def _patched_de_show_cal(self):
            try: self._last_known_date = self.get_date()
            except: self._last_known_date = None
            try: _orig_de_show_cal(self)
            except BaseException: pass
            if hasattr(self, '_calendar'):
                try: self._calendar.configure(select_on_nav=False)
                except: pass
        DateEntry._show_calendar = _patched_de_show_cal
except Exception:
    pass

# --- GLOBAL UTILITY FUNCTIONS MOVED TO utils/helpers.py ---


# 4) DateEntry.drop_down : tkcalendar 1.6.1+ 날짜박스 클릭 시 튕김 방지 (핵심 패치)
try:
    if hasattr(DateEntry, 'drop_down'):
        _orig_drop_down = DateEntry.drop_down
        def _patched_drop_down(self):
            try: self._last_known_date = self.get_date()
            except: self._last_known_date = None
            try:
                _orig_drop_down(self)
                if hasattr(self, '_calendar'):
                    try: self._calendar.configure(select_on_nav=False)
                    except: pass
            except BaseException:
                pass
        DateEntry.drop_down = _patched_drop_down
except Exception:
    pass

# 5) DateEntry._on_b1_press : 클릭 이벤트 처리 중 예외 방지
try:
    if hasattr(DateEntry, '_on_b1_press'):
        _orig_b1_press = DateEntry._on_b1_press
        def _patched_b1_press(self, event):
            try: _orig_b1_press(self, event)
            except BaseException: pass
        DateEntry._on_b1_press = _patched_b1_press
except Exception:
    pass

# 6) DateEntry._on_calendar_selection
try:
    if hasattr(DateEntry, '_on_calendar_selection'):
        _orig_de_on_sel = DateEntry._on_calendar_selection
        def _patched_de_on_sel(self, event):
            try:
                new_date = self._calendar.selection_get()
                if hasattr(self, '_last_known_date') and self._last_known_date == new_date:
                    pass
            except: pass
            _orig_de_on_sel(self, event)
        DateEntry._on_calendar_selection = _patched_de_on_sel
except Exception:
    pass

# 7) DateEntry._setup_style : 스타일 설정 중 update_idletasks 일시 차단
try:
    if hasattr(DateEntry, '_setup_style'):
        _orig_setup_style = DateEntry._setup_style
        def _patched_setup_style(self):
            _orig_update = self.update_idletasks
            self.update_idletasks = lambda: None
            try:
                _orig_setup_style(self)
            except BaseException:
                pass
            finally:
                self.update_idletasks = _orig_update
        DateEntry._setup_style = _patched_setup_style
except Exception:
    pass

# 8) DateEntry.update_idletasks : Python 3.14 BaseException 안전화 (클래스 레벨 영구 적용)
try:
    def _safe_de_update_idletasks(self):
        try:
            _tk_ref.Misc.update_idletasks(self)
        except BaseException:
            pass
    DateEntry.update_idletasks = _safe_de_update_idletasks
except Exception:
    pass

# 9) DateEntry._determine_downarrow_name : Configure/Map 이벤트 콜백 예외 방지
try:
    if hasattr(DateEntry, '_determine_downarrow_name'):
        _orig_det = DateEntry._determine_downarrow_name
        def _patched_det(self, event=None):
            try: _orig_det(self, event)
            except BaseException: pass
        DateEntry._determine_downarrow_name = _patched_det
except Exception:
    pass

# 10) Calendar._setup_style : ThemeChanged 시 예외 방지
try:
    if hasattr(Calendar, '_setup_style'):
        _orig_cal_setup_style = Calendar._setup_style
        def _patched_cal_setup_style(self, event=None):
            try: _orig_cal_setup_style(self, event)
            except BaseException: pass
        Calendar._setup_style = _patched_cal_setup_style
except Exception:
    pass

# 11) Calendar._prev_year / _next_year : 년도 이동
# [FIX] Exception 수준에서만 잡고, 날짜도 롤백.
try:
    if hasattr(Calendar, '_prev_year'):
        _orig_prev_year = Calendar._prev_year
        def _patched_prev_year(self):
            _saved = self._date
            try:
                _orig_prev_year(self)
            except Exception:
                self._date = _saved   # 실패 시 날짜 롤백
        Calendar._prev_year = _patched_prev_year
except Exception:
    pass

try:
    if hasattr(Calendar, '_next_year'):
        _orig_next_year = Calendar._next_year
        def _patched_next_year(self):
            _saved = self._date
            try:
                _orig_next_year(self)
            except Exception:
                self._date = _saved   # 실패 시 날짜 롤백
        Calendar._next_year = _patched_next_year
except Exception:
    pass

# 12) Calendar._prev_month / _next_month : 월 이동도 동일하게 안전화
try:
    if hasattr(Calendar, '_prev_month'):
        _orig_prev_month = Calendar._prev_month
        def _patched_prev_month(self):
            _saved = self._date
            try:
                _orig_prev_month(self)
            except Exception:
                self._date = _saved
        Calendar._prev_month = _patched_prev_month
except Exception:
    pass

try:
    if hasattr(Calendar, '_next_month'):
        _orig_next_month = Calendar._next_month
        def _patched_next_month(self):
            _saved = self._date
            try:
                _orig_next_month(self)
            except Exception:
                self._date = _saved
        Calendar._next_month = _patched_next_month
except Exception:
    pass

# 12.5) tk.Misc.destroy : Python 3.13+ 종료 시 TclError ("can't delete Tcl command") 예외 방지
try:
    _orig_misc_destroy = tk.Misc.destroy
    def _patched_misc_destroy(self):
        try:
            _orig_misc_destroy(self)
        except tk.TclError:
            pass
        except Exception:
            pass
    tk.Misc.destroy = _patched_misc_destroy
except Exception:
    pass

# 13) DateEntry._on_focus_out_cal : 년도/월 네비게이션 버튼 클릭 시 달력 닫힘 방지
# [근본 원인] 년도 버튼(_l_year/_r_year) 클릭 → Calendar에 <FocusOut> 발생
# → focus_get()이 버튼 위젯(DateEntry가 아닌 _top_cal 자식)을 반환
# → 기존 코드: "focus != self(DateEntry)" 이므로 else: _top_cal.withdraw() → 달력 닫힘
# → 사용자 입장: 년도 버튼 누를 때마다 달력이 꺼짐 ("화면 튕김")
# [수정] focus가 _top_cal 내부 위젯으로 이동한 경우 → 달력 유지
try:
    if hasattr(DateEntry, '_on_focus_out_cal'):
        _orig_on_focus_out_cal = DateEntry._on_focus_out_cal
        def _patched_on_focus_out_cal(self, event):
            fw = self.focus_get()
            if fw is not None:
                try:
                    # focus가 _top_cal 내부(년도·월 버튼 등)로 이동한 경우 → 유지
                    if str(fw).startswith(str(self._top_cal)):
                        return
                except Exception:
                    pass
            try:
                _orig_on_focus_out_cal(self, event)
            except BaseException:
                pass
        DateEntry._on_focus_out_cal = _patched_on_focus_out_cal
except Exception:
    pass

# --- Custom Draggable Messagebox Implementation ---

# Injection: Replace standard messagebox methods to enable draggable behavior globally
if not hasattr(messagebox, 'showerror_orig'):
    messagebox.showerror_orig = messagebox.showerror
    messagebox.showwarning_orig = messagebox.showwarning
    messagebox.showinfo_orig = messagebox.showinfo
    
    messagebox.showerror = DraggableMessagebox.showerror
    messagebox.showwarning = DraggableMessagebox.showwarning
    messagebox.showinfo = DraggableMessagebox.showinfo
















        
        # self.wait_window() removed from here to allow caller to set vars before blocking
        # self.wait_window() removed from here to allow caller to set vars before blocking

class MaterialManager:
    def __init__(self, root):
        self.root = root

        # 공정별 관경에 따른 검사길이 (m) 룩업 테이블
        self.SIZE_LENGTH = {
            '1100A': 3.511,  '1000A': 3.1919, '900A': 2.8727, '850A': 2.7131,
            '800A':  2.5535, '750A':  2.3939, '700A': 2.2343, '650A': 2.0747,
            '600A':  1.9151, '550A':  1.7555, '500A': 1.5959, '450A': 1.4363,
            '400A':  1.2767, '350A':  1.1172, '300A': 1.0006, '250A': 0.8401,
            '200A':  0.6795, '150A':  0.519,  '125A': 0.4392, '100A': 0.3591,
            '80A':   0.2799, '65A':   0.2397, '50A':  0.1901, '40A':  0.1527,
            '32A':   0.1341, '25A':   0.1068, '20A':  0.0855,
        }
        
        # High DPI awareness
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            try:
                ctypes.windll.user32.SetProcessDPIAware()
            except Exception:
                pass
                
        self.root.title("자재 및 소모품 관리 시스템 (Material Manager)")
        self.root.geometry("1600x900")
        try:
            self.root.state('zoomed') # Maximize on Windows
        except:
            pass
        
        # Configure overall style
        
        # Global Scroll Handler for all canvases in the app
        def _on_global_mousewheel(event):
            try:
                widget = event.widget.winfo_containing(event.x_root, event.y_root)
                if not widget: return
                
                # Do not interfere with native scrolling widgets
                if isinstance(widget, (tk.Text, ttk.Treeview, tk.Listbox)):
                    return
                    
                parent = widget
                while parent:
                    if isinstance(parent, tk.Canvas):
                        # Allow scrolling only if content is taller than the canvas
                        bbox = parent.bbox("all")
                        if bbox:
                            content_height = bbox[3] - bbox[1]
                            if content_height > parent.winfo_height():
                                parent.yview_scroll(int(-1 * (event.delta / 120)), "units")
                        return "break" # [FIX] Prevent scroll from reaching Combobox/Entry under mouse

                    if hasattr(parent, 'master') and parent.master:
                        parent = parent.master
                    else:
                        break
            except Exception:
                pass
            return "break" # Default to break if we are over a canvas-descendant to be safe
                
        self.root.bind_all("<MouseWheel>", _on_global_mousewheel)
        
        # [FIX] Globally disable MouseWheel on Combobox and Entry to prevent unintentional value changes during scrolling
        self.root.bind_class("TCombobox", "<MouseWheel>", lambda e: "break")
        self.root.bind_class("TEntry", "<MouseWheel>", lambda e: "break")
        
        self.style = ttk.Style()
        try:
            self.style.theme_use('clam') # Use 'clam' theme for better grid line visibility
        except:
            pass
            
        self.style.configure(".", font=('Malgun Gothic', 12))
        self.style.configure("Treeview.Heading", font=('Malgun Gothic', 12, 'bold'))
        self.style.configure("Treeview", font=('Malgun Gothic', 12), rowheight=35) # Increased row height for "boxed" look
        
        # Detect system background color for tk widgets (Canvas, Text)
        self.theme_bg = self.root.cget('bg')
        if not self.theme_bg or self.theme_bg == 'SystemButtonFace':
             # Fallback/Modern check for Windows
             try:
                 self.theme_bg = self.style.lookup('TFrame', 'background')
             except:
                 self.theme_bg = '#f0f0f0' # Standard Windows gray
        
        # Detachable Windows State (pop-up support)
        self.detached_windows = {}
        
        # Determine base directory and bundle directory for portability
        if getattr(sys, 'frozen', False):
            # If running as an executable
            self.app_dir = os.path.dirname(sys.executable)
            self.bundle_dir = getattr(sys, '_MEIPASS', self.app_dir)
            # For exe, use user's Documents folder for config to ensure write permissions
            documents_dir = os.path.join(os.path.expanduser('~'), 'Documents', 'MaterialManager', 'central')
            if not os.path.exists(documents_dir):
                os.makedirs(documents_dir, exist_ok=True)
            self.config_path = os.path.join(documents_dir, 'Material_Manager_Config.json')
        else:
            # 스크립트 실행 모드: src/ 우선, 없으면 ../data/ 탐색
            self.app_dir = os.path.dirname(os.path.abspath(__file__))
            self.bundle_dir = self.app_dir
            self.config_path = os.path.join(self.app_dir, 'Material_Manager_Config.json')
            _db_name = 'Material_Inventory.xlsx'
            _data_dir = os.path.join(os.path.dirname(self.app_dir), 'data')
            _candidates = [
                os.path.join(self.app_dir, _db_name),   # src/Material_Inventory.xlsx
                os.path.join(_data_dir, _db_name),      # data/Material_Inventory.xlsx
            ]
            self.db_path = next((p for p in _candidates if os.path.exists(p)), _candidates[0])

        if getattr(sys, 'frozen', False):
            # exe 실행 모드: exe와 같은 폴더에 DB 저장 (어디서 실행해도 동일)
            self.db_path = os.path.join(self.app_dir, 'Material_Inventory.xlsx')

        # db_path 확정
        
        self.sites = [] # Initialize site list
        self.daily_units = ['EA', 'CAN', 'SET', 'KG', 'M', '매', 'I/D', 'P,M,I/D', 'M,I/D', 'Point', 'Meter', 'Inch', 'Dia']
        self.users = [
            "부장 주진철", "대리 우명광", "주임 김진환", "계장 장승대", "주임 김성렬", "부장 박광복", "과장 주영광"
        ] # Initialize worker/name list
        self.warehouses = [] # Initialize warehouse list
        self.equipments = [] # Initialize equipment list
        self.vehicles = [] # Initialize vehicle list
        self.vehicle_inspections = {} # [NEW] Track active inspection widgets
        self.vehicle_boxes = [] # [NEW] Track active inspection widgets for automated save
        self.companies = [] # Initialize companies list
        self.worktimes = [] # Initialize worktimes list
        self.ot_times = [] # Initialize ot_times list
        self.test_items = [] # [NEW] Initialize test_items
        self.applied_codes = [] # [NEW] Initialize applied_codes
        
        # [NEW] Pre-initialize autocomplete/display lists to prevent AttributeError during load_data/refresh_filters
        self.materials_display_list = []
        self.co_code_list = []
        self.eq_code_list = []
        self.item_name_list = []
        self.class_list = []
        self.spec_list = []
        self.unit_list = []
        self.mfr_list = []
        self.origin_list = []
        self.sn_list = []
        self.model_list = []
        self.supplier_list = []
        self.equipment_suggestions = []
        self.test_methods = [" ", "RT", "PAUT", "UT", "MT", "PT", "PMI", "ETC"]  # [NEW] Initialize test methods
        
        # [NEW] Centralized NDT Consumable Definitions
        self.ndt_groups = {
            'PT약품': ['세척제', '침투제', '현상제', '형광침투제'],
            'MT약품': ['백색페인트', '흑색자분', '형광자분']
        }
        self.ndt_materials_all = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
        
        self.budget_sites = []  # [NEW] 공사실행예산서 전용 현장 목록 (budget_df 현장만)
        self.hidden_sites = ["초안", "롯데현장"]  # [NEW] Default hidden sites as requested
        
        self.load_data()
        
        # Centralized list of Carestream films for suggestions
        self.carestream_films = [
            "Carestream AA400-3⅓*12\"",
            "Carestream AA400-3⅓*17\"", 
            "Carestream AA400-4½*12\"",
            "Carestream AA400-10*12\"",
            "Carestream M100-3⅓*12\"",
            "Carestream M100-10*12\"",
            "Carestream M100-14*17\"",
            "Carestream MX125-3⅓*6\"",
            "Carestream MX125-3⅓*12\"",
            "Carestream MX125-4½*12\"",
            "Carestream MX125-10*12\"",
            "Carestream MX125-14*17\"",
            "Carestream T200-3⅓*12\"",
            "Carestream T200-3⅓*17\"",
            "Carestream T200-4½*12\"",
            "Carestream T200-10*12\"",
            "Carestream T200-14*17\""
        ]
        
        # Core draggable widget keys that should never be hidden
        self.CORE_DRAGGABLE_KEYS = [
            'form_box_geometry', 
            'ndt_usage_box_geometry', 
            'rtk_usage_box_geometry', 
            'save_btn_geometry', 
            'workers_box_geometry'
        ]
        
        # Registry for draggable items {config_key: widget_instance}
        self.draggable_items = {}
        self.memos = {} # key -> {'container': container, 'text_widget': text, 'title_entry': entry}
        self.checklists = {} # key -> {'container': container, 'title_entry': entry, 'item_frame': frame, 'items': []}
        self.vehicle_inspections = {} # key -> widget
        self.layout_locked = False
        self.daily_usage_sash_locked = False
        self._last_motion_time = 0 # For performance throttling
        self.is_ready = False  # Suppress saves until fully loaded

        # [CRITICAL FIX] Pre-load configuration synchronously to ensure UI starts in correct state
        self.preload_config_locks()

        self.create_widgets()
        self.update_registration_combos()
        
        # Enable keyboard navigation
        self.setup_keyboard_shortcuts()

        # NDT 변수 초기화 및 추적 바인딩 추가
        self._init_ndt_vars_and_traces()

    def _init_ndt_vars_and_traces(self):
        """NDT 관련 변수(관경, 조인트수)를 생성하고 자동계산 이벤트를 바인딩합니다."""
        # 동적 바인딩에 대비해 명시적으로 StringVar 초기화
        for var_name in ['ndt_report_pipe_var', 'ndt_ori_joint_var', 'ndt_rep_joint_var', 'ndt_ori_qty_var', 'ndt_rep_qty_var', 'ndt_inspection_type_var']:
            if not hasattr(self, var_name):
                setattr(self, var_name, tk.StringVar(value=""))

        # 콜백 함수: 관경과 조인트수가 변경되면 수량을 자동계산
        def on_ndt_change(*args):
            try:
                if not hasattr(self, 'cb_daily_test_method'): return
                method = self.cb_daily_test_method.get().strip().upper()
                
                # PAUT, PT, MT 공정만 자동 계산
                if not any(x in method for x in ['PAUT', 'PT', 'MT']): return
                
                pipe_val = getattr(self, 'ndt_report_pipe_var').get().strip().upper()
                
                # 숫자만 입력된 경우 'A'를 붙여서 처리 (예: "600" -> "600A")
                if pipe_val and pipe_val.isdigit():
                    pipe_val += 'A'
                    
                if pipe_val not in self.SIZE_LENGTH: return
                
                length = self.SIZE_LENGTH[pipe_val]
                
                # 원본 조인트수
                ori_j = getattr(self, 'ndt_ori_joint_var').get().strip()
                # 재촬영 조인트수
                rep_j = getattr(self, 'ndt_rep_joint_var').get().strip()
                
                total_joints = 0.0
                
                if ori_j:
                    try:
                        joints = float(ori_j.replace(',', ''))
                        total_joints += joints
                        getattr(self, 'ndt_ori_qty_var').set(f"{joints * length:.4f}")
                    except ValueError: pass
                    
                if rep_j:
                    try:
                        joints = float(rep_j.replace(',', ''))
                        total_joints += joints
                        getattr(self, 'ndt_rep_qty_var').set(f"{joints * length:.4f}")
                    except ValueError: pass
                    
                # 메인 폼의 '수량' 칸(ent_daily_test_amount)에 실시간 자동 기입
                if total_joints > 0 and hasattr(self, 'ent_daily_test_amount'):
                    self.ent_daily_test_amount.delete(0, tk.END)
                    self.ent_daily_test_amount.insert(0, f"{total_joints * length:.4f}")
                    
            except Exception as e:
                print(f"NDT auto-calc error: {e}")

        # 변수에 추적(trace) 추가
        getattr(self, 'ndt_report_pipe_var').trace_add('write', on_ndt_change)
        getattr(self, 'ndt_ori_joint_var').trace_add('write', on_ndt_change)
        getattr(self, 'ndt_rep_joint_var').trace_add('write', on_ndt_change)

    def preload_config_locks(self):
        """Pre-load critical lock states synchronously before UI creation"""
        try:
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.layout_locked = config.get('layout_locked', False)
                    self.daily_usage_sash_locked = config.get('daily_usage_sash_locked', False)
                    print(f"DEBUG: Pre-loaded locks - Layout: {self.layout_locked}, Sash: {self.daily_usage_sash_locked}")
        except Exception as e:
            print(f"DEBUG: Failed to pre-load locks: {e}")

    def _ensure_canvas_scroll_region(self):
        """Update canvas scroll region based on content height and width (ensures full visibility)"""
        try:
            if hasattr(self, 'entry_canvas') and self.entry_canvas:
                # Remove update_idletasks() to prevent infinite <Configure> loops 
                # Instead, use an after() callback to defer the bbox calculation if needed, or just calculate as-is
                
                def _do_update():
                    if not self.entry_canvas.winfo_exists(): return
                    
                    # Dynamically adjust the canvas window height to accommodate new/removed widgets
                    if hasattr(self, 'entry_inner_frame') and hasattr(self, 'entry_canvas_window'):
                        req_h = self.entry_inner_frame.winfo_reqheight()
                        canvas_h = self.entry_canvas.winfo_height()
                        target_h = max(req_h, canvas_h)
                        self.entry_canvas.itemconfig(self.entry_canvas_window, height=target_h)
                        
                    bbox = self.entry_canvas.bbox("all")
                    if bbox:
                        canvas_h = self.entry_canvas.winfo_height()
                        canvas_w = self.entry_canvas.winfo_width()
                        bottom = max(bbox[3], canvas_h)
                        right = max(bbox[2], canvas_w, 1100)
                        self.entry_canvas.configure(scrollregion=(0, 0, right, bottom))
                        
                # Debounce the update to prevent recursive looping
                if hasattr(self, '_scroll_update_id'):
                    self.entry_canvas.after_cancel(self._scroll_update_id)
                self._scroll_update_id = self.entry_canvas.after(50, _do_update)
                
        except Exception as e:
            print(f"DEBUG: Scroll region update error: {e}")

    def _on_daily_usage_sash_changed(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import _on_daily_usage_sash_changed_impl
        return _on_daily_usage_sash_changed_impl(self, *args, **kwargs)

    def toggle_resolution_lock(self):
        """Toggle window resolution lock"""
        try:
            self.resolution_locked = not self.resolution_locked
            
            if self.resolution_locked:
                self.locked_width = self.root.winfo_width()
                self.locked_height = self.root.winfo_height()
                self.root.resizable(False, False)
                if hasattr(self, 'btn_resolution_lock'):
                    self.btn_resolution_lock.config(text="🔒 해상도 고정됨")
                
                if not hasattr(self, 'tab_config'):
                    self.tab_config = {}
                self.tab_config['resolution_locked'] = True
                self.tab_config['locked_width'] = self.locked_width
                self.tab_config['locked_height'] = self.locked_height
                print(f"Resolution locked at: {self.locked_width}x{self.locked_height}")
            else:
                self.root.resizable(True, True)
                if hasattr(self, 'btn_resolution_lock'):
                    self.btn_resolution_lock.config(text="🔓 해상도 고정")
                
                if not hasattr(self, 'tab_config'):
                    self.tab_config = {}
                self.tab_config['resolution_locked'] = False
                print("Resolution unlocked")
                
            self.save_tab_config(force=True)
        except Exception as e:
            print(f"Error toggling resolution lock: {e}")

    # Removed force_save_config - using save_tab_config(force=True) instead

    def extract_sn_from_model(self, model_name, current_sn):
        """Extract SN from model name if 'S/N.' is present and update current_sn if it's empty"""
        if pd.isna(model_name):
            return model_name, current_sn
        
        model_str = str(model_name)
        if 'S/N.' in model_str:
            parts = model_str.split('S/N.', 1)
            new_model = parts[0].strip()
            # Remove trailing underscore or dot if present before S/N.
            if new_model.endswith('_') or new_model.endswith('.'):
                new_model = new_model[:-1].strip()
            
            extracted_sn = parts[1].strip()
            
            # Use extracted SN if current SN is empty or matches part of it
            if not str(current_sn).strip() or str(current_sn) == 'nan':
                return new_model, extracted_sn
            else:
                # If current SN exists, we still clean up the model name
                return new_model, current_sn
        
        return model_name, current_sn

    def _sync_dataframe_schema(self, *args, **kwargs):
        from site_apps.central.src.models.material_model import _sync_dataframe_schema_impl
        return _sync_dataframe_schema_impl(self, *args, **kwargs)

    def load_data(self, *args, **kwargs):
        if "export" in "load_data" or "excel" in "load_data":
            from site_apps.central.src.services.excel_exporter import load_data_impl
        else:
            from site_apps.central.src.services.data_loader import load_data_impl
        return load_data_impl(self, *args, **kwargs)


    
    def migrate_old_schema(self):
        """Migrate data from old schema to new schema"""
        old_df = self.materials_df.copy()
        self.materials_df = pd.DataFrame(columns=[
            'MaterialID', '회사코드', '관리품번', '품목명', 'SN', '창고',
            '모델명', '규격', '품목군코드', '제조사', '제조국', 
            '가격', '원가', '관리단위', '수량'
        ])
        
        # Headers are already cleaned by the global cleanup in load_data/init

        for _, row in old_df.iterrows():
            new_row = {
                'MaterialID': row.get('MaterialID', ''),
                '회사코드': '',
                '관리품번': row.get('Equipment Code', ''),
                '품목명': row.get('Item Name', row.get('Name', '')),
                'SN': row.get('SN', ''),
                '창고': '',
                '모델명': '',
                '규격': row.get('Specification', ''),
                '품목군코드': '',
                '제조사': row.get('Manufacturer', ''),
                '제조국': '',
                '가격': 0,
                '관리단위': row.get('Unit', 'EA'),
                '수량': row.get('Current Stock', row.get('Initial Stock', 0))
            }
            self.materials_df = pd.concat([self.materials_df, pd.DataFrame([new_row])], ignore_index=True)
        
        self.save_data()
        messagebox.showinfo("마이그레이션 완료", "기존 데이터가 새로운 형식으로 변환되었습니다.")

    def register_new_material(self, name, model='', sn='', **kwargs):
        """Helper to register a new material in the master list summerly"""
        try:
            # Robustly calculate NEW ID even if mixed types or NaNs exist
            valid_ids = pd.to_numeric(self.materials_df['MaterialID'], errors='coerce').dropna()
            new_id = int(valid_ids.max() + 1) if not valid_ids.empty else 10001
        except:
            new_id = 10001 # Fallback
        
        # Determine defaults
        new_mat = {
            'MaterialID': new_id,
            '회사코드': kwargs.get('co_code', ''),
            '관리품번': '',
            '품목명': name,
            'SN': sn,
            '창고': kwargs.get('warehouse', ''),
            '모델명': model or name,
            '규격': '',
            '품목군코드': '',
            '공급업체': '',
            '제조사': '',
            '제조국': '',
            '가격': 0,
            '원가': 0,
            '관리단위': 'EA',
            '수량': 0,
            '재고하한': 10,
            'Active': 1
        }
        # Update with any explicit kwargs
        for k, v in kwargs.items():
            if k in new_mat: new_mat[k] = v
            
        self.materials_df = pd.concat([self.materials_df, pd.DataFrame([new_mat])], ignore_index=True)
        self.save_data()
        self.update_material_combo()
        return new_id
        
        self.save_data()
        messagebox.showinfo("마이그레이션 완료", "기존 데이터가 새로운 형식으로 변환되었습니다.")
    
    def enable_autocomplete(self, combobox, values_list_attr=None, values_list=None, prefix_list=None):
        """Standardize to Native Dropdown: Filter values but use standard system dropdown."""
        if not hasattr(self, '_autocomplete_timers'):
            self._autocomplete_timers = {}

        def perform_filter(event_widget, force_all=False):
            """Filter combobox values based on typed text."""
            typed = str(event_widget.get()).strip()
            widget_name = str(event_widget)
            
            if not typed and not force_all:
                return  # Don't filter on empty input unless forced
            
            if values_list_attr:
                base_values = getattr(self, values_list_attr, [])
            elif values_list is not None:
                base_values = values_list
            else:
                base_values = []
            
            all_values = (prefix_list + base_values) if prefix_list else base_values
            
            if force_all or not typed:
                new_values = all_values
            elif prefix_list and typed in prefix_list:
                new_values = all_values
            else:
                new_values = [v for v in all_values if typed.lower() in str(v).lower()]
            
            print(f"[FILTER] new_values count: {len(new_values)}, changed: {list(event_widget['values']) != list(new_values)}")
            
            # Update values if changed
            if list(event_widget['values']) != list(new_values):
                event_widget['values'] = new_values
            
            # Auto-open dropdown after typing (only if not empty and has focus)
            if not force_all and typed and len(new_values) > 0:
                # Cancel any pending open
                if hasattr(event_widget, '_dropdown_timer'):
                    event_widget.after_cancel(event_widget._dropdown_timer)
                
                def open_dropdown():
                    try:
                        # [STABILITY] Check if it was recently selected to avoid "sticky" dropdown
                        if getattr(event_widget, '_just_selected', False):
                            return

                        # Only if still has focus and NOT already open (heuristic)
                        if event_widget.focus_get() == event_widget:
                            pos = event_widget.index(tk.INSERT)
                            event_widget.event_generate('<Alt-Down>')
                            # Keep typing position
                            event_widget.after(50, lambda: event_widget.icursor(pos))
                    except Exception:
                        pass
                
                event_widget._dropdown_timer = event_widget.after(400, open_dropdown)


        def on_keyrelease(event):
            """Filter values as user types. Down arrow or click ▼ to see filtered list."""
            # Cancel any pending dropdown open while typing
            if hasattr(combobox, '_dropdown_timer'):
                combobox.after_cancel(combobox._dropdown_timer)
                delattr(combobox, '_dropdown_timer')
            
            if event.keysym in ("Left", "Right", "Up", "Down", "Return", "Escape", "Tab", "Shift_L", "Shift_R", "Control_L", "Control_R"):
                return
            
            # Use a timer to wait for typing to stop before filtering
            if combobox in self._autocomplete_timers:
                self.root.after_cancel(self._autocomplete_timers[combobox])
            
            def delayed_filter():
                try:
                    # Poll for value to handle Korean IME composition
                    max_polls = 20
                    poll_count = [0]
                    last_val = ['']
                    
                    def poll_for_value():
                        try:
                            current_val = combobox.get()
                            if not last_val[0] and current_val:
                                perform_filter(combobox, force_all=False)
                                return
                            if not current_val and poll_count[0] < max_polls:
                                last_val[0] = current_val
                                poll_count[0] += 1
                                combobox.after(50, poll_for_value)
                                return
                            if current_val:
                                perform_filter(combobox, force_all=False)
                            else:
                                perform_filter(combobox, force_all=True)
                        except Exception: pass
                    
                    poll_for_value()
                except Exception: pass
                
            self._autocomplete_timers[combobox] = self.root.after(200, delayed_filter)

        combobox.bind('<KeyRelease>', on_keyrelease, add='+')

        # [NEW] Guard against "sticky" dropdown after selection
        def _on_selected(e=None):
            combobox._just_selected = True
            def _reset(): combobox._just_selected = False
            combobox.after(500, _reset)
            
        combobox.bind('<<ComboboxSelected>>', _on_selected, add='+')
        
        # [NEW] Also bind to KeyPress for Korean input composition support
        def on_keypress(event):
            """Handle KeyPress for Korean input composition."""
            print(f"[KEYPRESS] keysym={event.keysym}, char='{event.char}'")
            # For Korean input, schedule filter on any keypress too
            if event.keysym not in ("Left", "Right", "Up", "Down", "Return", "Escape", "Tab", "Shift_L", "Shift_R", "Control_L", "Control_R"):
                # Cancel existing timer
                if combobox in self._autocomplete_timers:
                    self.root.after_cancel(self._autocomplete_timers[combobox])
                
                # Also cancel dropdown timer
                if hasattr(combobox, '_dropdown_timer'):
                    combobox.after_cancel(combobox._dropdown_timer)
                    delattr(combobox, '_dropdown_timer')
                
                # Schedule filter with polling for Korean IME composition
                def delayed_filter_korean():
                    try:
                        # Poll for Korean IME composition - check every 50ms for up to 1 second
                        max_polls = 20
                        poll_count = [0]
                        last_val = ['']
                        
                        def poll_for_value():
                            try:
                                current_val = combobox.get()
                                print(f"[KOREAN-POLL] count={poll_count[0]}, value='{current_val}'")
                                
                                # If value appeared, filter immediately
                                if not last_val[0] and current_val:
                                    print(f"[KOREAN-POLL] Value appeared: '{current_val}'")
                                    perform_filter(combobox, force_all=False)
                                    return
                                
                                # Keep polling while empty
                                if not current_val and poll_count[0] < max_polls:
                                    last_val[0] = current_val
                                    poll_count[0] += 1
                                    combobox.after(50, poll_for_value)
                                    return
                                
                                # Done polling
                                if current_val:
                                    perform_filter(combobox, force_all=False)
                                else:
                                    # [KOREAN IME] Show all values
                                    print(f"[KOREAN-POLL] Timeout, showing all values")
                                    perform_filter(combobox, force_all=True)
                            except Exception as e:
                                print(f"[KOREAN-POLL] Error: {e}")
                        
                        poll_for_value()
                    except Exception as e:
                        print(f"[KOREAN] Error: {e}")
                
                timer_id = self.root.after(400, delayed_filter_korean)  # 400ms for Korean IME
                self._autocomplete_timers[combobox] = timer_id
        
        combobox.bind('<KeyPress>', on_keypress, add=True)

        def on_focus_in(event):
            try:
                combobox.selection_range(0, tk.END)
            except: pass
            perform_filter(combobox, force_all=True)
            
        combobox.bind('<FocusIn>', on_focus_in, add=True)
        
        # Click to show full list
        combobox.bind('<Button-1>', lambda e: perform_filter(combobox, force_all=True), add=True)


    def apply_autocomplete_to_all_comboboxes(self):
        """Map specific comboboxes to their data lists for autocomplete"""
        # Mapping: {widget_attr_name: values_list_attr_name}
        mappings = {
            'cb_material': 'materials_display_list',
            'cb_trans_filter_mat': 'materials_display_list',
            'cb_trans_filter_site': 'sites',
            'cb_budget_site': 'budget_sites',
            'cb_daily_filter_worker': 'users',
            'cb_daily_filter_vehicle': 'vehicles',
            'cb_daily_test_method': 'test_methods',  # Test method list
            'cb_daily_co_code': 'co_code_list',  # [NEW] Company code for MT/PT
            'cb_trans_filter_vehicle': 'vehicles',
            'cb_sales_filter_site': 'sites',
            'cb_filter_co': 'co_code_list',
            'cb_filter_class': 'class_list',
            'cb_filter_mfr': 'mfr_list',
            'cb_filter_name': 'consumable_display_list', # [NEW] Use consumable-only list for Stock View
            'cb_filter_sn': 'sn_list',
            'cb_filter_model': 'model_list',
            'cb_filter_eq': 'eq_code_list',
            'cb_co_code': 'co_code_list',
            'cb_eq_code': 'eq_code_list',
            'cb_item_name': 'item_name_list',
            'cb_model': 'model_list',
            'cb_class': 'class_list',
            'cb_spec': 'spec_list',
            'cb_unit': 'unit_list',
            'cb_supplier': 'supplier_list',
            'cb_mfr': 'mfr_list',
            'cb_origin': 'origin_list',
            'cb_trans_site': 'sites',
            'cb_warehouse': 'warehouses',
            'ent_user': 'users',
            # [NEW] Additional comboboxes for autocomplete
            'cb_name': 'users',  # Worker name
            'ent_worktime': 'worktimes',  # Work time
            'cb_vehicle_info': 'vehicles',  # Vehicle info
        }
        
        for widget_attr, list_attr in mappings.items():
            if hasattr(self, widget_attr):
                widget = getattr(self, widget_attr)
                if isinstance(widget, ttk.Combobox):
                    # For filtering, we might want to include '전체' dynamically
                    current_values = list(widget['values'])
                    if '전체' in current_values:
                        self.enable_autocomplete(widget, values_list_attr=list_attr, prefix_list=['전체'])
                    else:
                        self.enable_autocomplete(widget, values_list_attr=list_attr)

    def _safe_format_datetime(self, val, format_str='%Y-%m-%d %H:%M'):
        if pd.isna(val) or val is None or str(val).strip().lower() in ['nan', 'none', '']:
            return ""
        try:
            return pd.to_datetime(val).strftime(format_str)
        except:
            return str(val)

    def clean_nan(self, *args, **kwargs):
        from site_apps.central.src.utils.helpers import clean_nan_impl
        return clean_nan_impl(self, *args, **kwargs)

    def normalize_id(self, val):
        """Robustly normalize IDs: handle NaN, trailing .0, and whitespace."""
        if pd.isna(val) or val == '' or str(val).lower() == 'nan': return ""
        s = str(val).strip()
        if s.endswith('.0'): s = s[:-2]
        return s

    def save_data(self, *args, **kwargs):
        if "export" in "save_data" or "excel" in "save_data":
            from site_apps.central.src.services.excel_exporter import save_data_impl
        else:
            from site_apps.central.src.services.data_loader import save_data_impl
        return save_data_impl(self, *args, **kwargs)
            
    def get_base_salaries(self, *args, **kwargs):
        from site_apps.central.src.models.worker_model import get_base_salaries_impl
        return get_base_salaries_impl(self, *args, **kwargs)

    def get_material_defaults(self, *args, **kwargs):
        from site_apps.central.src.models.material_model import get_material_defaults_impl
        return get_material_defaults_impl(self, *args, **kwargs)

    def get_expense_defaults(self):
        """Extract site expense defaults from settings_df"""
        if not hasattr(self, 'settings_df') or self.settings_df.empty:
            return [
                ("차량유지비", "주유, 수리, 통행, 주차 등", "N/A", 1, "일", 5000),
                ("소모품비", "장갑,일회용 작업복외", "N/A", 1, "일", 500),
                ("복리후생비", "생수, 음료 외 기타", "N/A", 1, "일", 1667),
                ("Se-175", "방사성동위원소 구매", "N/A", 1, "일", 47619)
            ]
        df = self.settings_df[self.settings_df['Category'] == 'Expense']
        if df.empty:
            return [
                ("차량유지비", "주유, 수리, 통행, 주차 등", "N/A", 1, "일", 5000),
                ("소모품비", "장갑,일회용 작업복외", "N/A", 1, "일", 500),
                ("복리후생비", "생수, 음료 외 기타", "N/A", 1, "일", 1667),
                ("Se-175", "방사성동위원소 구매", "N/A", 1, "일", 47619)
            ]
        # Return in (cat, cont, ppl, qty, unit, price) format as expected by _add_row_s1
        result = []
        for x in df[['Name', 'Spec', 'Unit', 'Rate']].values:
            name = x[0]
            spec = x[1]
            unit = x[2]
            rate = x[3]
            
            # 강제로 단가 및 규격 덮어쓰기
            if name == '차량유지비':
                unit = '일'
                rate = 5000
            elif name == '소모품비':
                unit = '일'
                rate = 500
            elif name == '복리후생비':
                unit = '일'
                rate = 1667
            elif name == 'Se-175':
                unit = '일'
                rate = 47619
                
            result.append((name, spec, "N/A", 1, unit, rate))
        return result

    def get_outsource_defaults(self):
        """Extract outsource defaults from settings_df"""
        if not hasattr(self, 'settings_df') or self.settings_df.empty:
            return [("케이엔디이", "방사선투과검사", 0, 15000)]
        df = self.settings_df[self.settings_df['Category'] == 'Outsource']
        if df.empty:
            return [("케이엔디이", "방사선투과검사", 0, 15000)]
        return [tuple(x) for x in df[['Name', 'Spec', 'Unit', 'Rate']].values]

    def create_widgets(self):
        # Notebook for Tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill='both', padx=10, pady=10)
        
        # Tab 1: Current Stock
        self.tab_stock = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_stock, text='현재 재고 현황')
        self.setup_stock_tab()
        
        # Tab 2: Register/Transaction
        self.tab_inout = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_inout, text='입출고 관리')
        self.setup_inout_tab()
        
        # Tab 3: Daily Work Log (New)
        self.tab_daily_work_log = DailyWorkLogTab(self.notebook)
        self.tab_daily_work_log.main_app = self
        self.notebook.add(self.tab_daily_work_log, text='작업/감독일보')

        # Tab 4: Import/Export
        self.tab_import = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_import, text='데이터 가져오기/내보내기')
        self.setup_import_tab()
        
        # Tab 5: Monthly Usage Entry
        self.tab_monthly_usage = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_monthly_usage, text='월별 집계')
        self.setup_monthly_usage_tab()
        
        # Tab 6: Daily Usage Entry by Site
        self.tab_daily_usage = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_daily_usage, text='현장별 일일 사용량 기입')
        self.setup_daily_usage_tab()
        
        # Tab 7: Daily Usage Query
        self.tab_daily_usage_query = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_daily_usage_query, text='현장 일일기록 조회 및 관리')
        self.setup_daily_usage_query_tab()
        

        # Tab 8: Project Execution Budget (New)
        self.tab_budget = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_budget, text='공사실행예산서')
        self.setup_budget_tab()
        
        # Tab 9: NDT Billing Calculator
        self.tab_ndt_billing = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_ndt_billing, text='기성 정산 (NDT)')
        self.setup_ndt_billing_tab()
        
        # Select default tab instantly to prevent flicker
        self.notebook.select(self.tab_daily_usage)
        
        # Initial view update (Move here to ensure all tabs are ready before refresh)
        self.refresh_inquiry_filters()
        self.update_daily_usage_view()
        
        # Bind tab events (Drag and drop reordering)
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        self.notebook.bind("<Button-1>", self.on_tab_drag_start)
        self.notebook.bind("<B1-Motion>", self.on_tab_drag)
        self.notebook.bind("<ButtonRelease-1>", self.on_tab_drag_end)
        self.notebook.bind("<Button-3>", self.show_tab_context_menu)
        
        # Bind main window resize to maintain sash ratios
        self.root.bind("<Configure>", self._on_main_window_resize)
        
        # Load and restore all configuration (geometry, locks, sashes, draggable items)
        self.root.after(100, self.load_tab_config)
        
        # Apply autocomplete to all comboboxes
        self.root.after(200, self.apply_autocomplete_to_all_comboboxes)
        
        # Save tab config on window close
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    def setup_stock_tab(self, *args, **kwargs):
        from site_apps.central.src.views.stock_management_view import setup_stock_tab_impl
        return setup_stock_tab_impl(self, *args, **kwargs)
    
    def show_low_stock(self):
        """Show items with low stock (less than their specific reorder point)"""
        low_stock_items = []
        for _, mat in self.materials_df.iterrows():
            if mat.get('Active', 1) == 0:
                continue
                
            current = self.calculate_current_stock(mat['MaterialID'])
            reorder_point = mat.get('재고하한', 10)
            if pd.isna(reorder_point) or reorder_point <= 0:
                reorder_point = 10 # Default fallback
                
            if pd.notna(current) and current < reorder_point:
                low_stock_items.append((mat.get('품목명', ''), current, reorder_point))
        
        if not low_stock_items:
            messagebox.showinfo("재고 알림", "수량이 10개 미만인 항목이 없습니다.")
        else:
            msg = "다음 항목들의 재고가 부족합니다:\n\n"
            for item, current, reorder in low_stock_items:
                msg += f"• {item}: 현재 {current:g} (필요 수준: {reorder:g})\n"
            messagebox.showwarning("재고 부족", msg)
    
    def select_all_stock(self):
        """Select all items in the stock treeview"""
        all_items = self.stock_tree.get_children()
        self.stock_tree.selection_set(all_items)

    def delete_selected_material(self):
        """선택된 자재 항목들을 재고에서 삭제"""
        selected_items = self.stock_tree.selection()
        
        if not selected_items:
            messagebox.showwarning("선택 오류", "삭제할 품목을 선택해주세요.")
            return
        
        # Confirm deletion
        confirm = messagebox.askyesno("삭제 확인", f"선택한 {len(selected_items)}개의 품목을 재고에서 영구히 삭제하시겠습니까?\n이 작업은 되돌릴 수 없습니다.")
        
        if not confirm:
            return
            
        # Get MaterialIDs to delete
        mat_ids_to_remove = []
        for item in selected_items:
            values = self.stock_tree.item(item, 'values')
            if values:
                # Ensure we match the type of MaterialID in the dataframe
                # [UX IMPROVEMENT] Get hidden MaterialID from the end of values
                mat_ids_to_remove.append(type(self.materials_df['MaterialID'].iloc[0])(values[-1]))
        
        # Soft delete: Set Active=0 instead of removing from materials_df
        if 'Active' not in self.materials_df.columns:
            self.materials_df['Active'] = 1
            
        # Standardize Active column to numeric to avoid pandas TypeError
        self.materials_df['Active'] = pd.to_numeric(self.materials_df['Active'], errors='coerce').fillna(1)
        
        initial_count = len(self.materials_df[self.materials_df['Active'] != 0])
        
        # Apply deletion (set to 0)
        mask = self.materials_df['MaterialID'].isin(mat_ids_to_remove)
        self.materials_df.loc[mask, 'Active'] = 0
        
        final_count = len(self.materials_df[self.materials_df['Active'] != 0])
        removed_count = initial_count - final_count
        
        if removed_count > 0:
            # Save data and update views
            self.save_data()
            self.update_stock_view()
            self.update_material_combo()
            
            # Optional: Clear transactions related to these materials?
            # For now, let's keep them for history unless explicitly asked.
            
            messagebox.showinfo("완료", f"{removed_count}개의 품목이 삭제되었습니다.")
        else:
            messagebox.showwarning("실패", "데이터프레임에서 항목을 삭제하지 못했습니다.")

    def reset_stock_filters(self):
        """Reset all stock filters to default"""
        self.cb_filter_co.set("전체")
        self.cb_filter_class.set("전체")
        self.cb_filter_mfr.set("전체")
        self.cb_filter_name.set("전체")
        self.cb_filter_sn.set("전체")
        self.cb_filter_model.set("전체")
        self.cb_filter_eq.set("전체")
        self.search_var.set("")
        self.update_stock_view()

    def open_edit_material_dialog(self):
        """Open a dialog to edit the selected material"""
        selection = self.stock_tree.selection()
        if not selection:
            messagebox.showwarning("선택 오류", "수정할 자재를 선택해주세요.")
            return
            
        item = self.stock_tree.item(selection[0])
        mat_values = item['values']
        # [UX IMPROVEMENT] Get hidden MaterialID from the end of values
        mat_id = mat_values[-1]
        
        # Get full material data from DF (Robust matching)
        target_id_str = self.normalize_id(mat_id)
        mask = self.materials_df['MaterialID'].apply(self.normalize_id) == target_id_str
        matches = self.materials_df[mask]
        
        if matches.empty:
            messagebox.showerror("오류", f"자재 ID '{mat_id}'를 데이터베이스에서 찾을 수 없습니다.")
            edit_win.destroy()
            return
            
        mat_data = matches.iloc[0]
        
        # Create Edit Dialog
        edit_win = tk.Toplevel(self.root)
        edit_win.title("자재 정보 수정")
        edit_win.geometry("550x750")
        edit_win.transient(self.root)
        edit_win.grab_set()
        
        # [NEW] Scrollable area for many fields (User request: button visibility)
        outer_frame = ttk.Frame(edit_win)
        outer_frame.pack(fill='both', expand=True)
        
        canvas = tk.Canvas(outer_frame, highlightthickness=0, bg=getattr(self, 'theme_bg', '#f0f0f0'))
        scrollbar = ttk.Scrollbar(outer_frame, orient="vertical", command=canvas.yview)
        main_frame = ttk.Frame(canvas, padding=20)
        
        main_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=main_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Enable mousewheel scrolling (Safer local binding)
        def _on_mousewheel(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except: pass
        edit_win.bind("<MouseWheel>", _on_mousewheel)
        
        fields = [
            ('회사코드', '회사코드'),
            ('관리품번', '관리품번'),
            ('품목명', '품목명'),
            ('SN', 'SN'),
            ('창고', '창고'),
            ('모델명', '모델명'),
            ('규격', '규격'),
            ('품목군코드', '품목군코드'),
            ('공급업체', '공급업체'),
            ('제조사', '제조사'),
            ('제조국', '제조국'),
            ('가격', '가격'),
            ('원가', '원가'),
            ('관리단위', '관리단위'),
            ('수량', '수량'),
            ('재고하한', '재고하한')
        ]
        
        entries = {}
        for i, (label_text, col_name) in enumerate(fields):
            ttk.Label(main_frame, text=f"{label_text}:").grid(row=i, column=0, padx=5, pady=5, sticky='w')
            
            # Using Combobox for some fields to maintain consistency
            if col_name in ['회사코드', '품목군코드', '제조사', '제조국', '관리단위']:
                ent = ttk.Combobox(main_frame, width=35)
                # [UNIFIED_UNIT] Populate values from managed list for units, others from data
                if col_name == '관리단위':
                    ent['values'] = getattr(self, 'unit_list', self.daily_units)
                elif col_name in self.materials_df.columns:
                    unique_vals = sorted(self.materials_df[col_name].dropna().unique().tolist())
                    ent['values'] = unique_vals
            else:
                ent = ttk.Entry(main_frame, width=38)
                
            ent.grid(row=i, column=1, padx=5, pady=5)
            
            # Pre-fill value
            val = mat_data.get(col_name, '')
            if pd.isna(val): val = ''
            ent.insert(0, str(val))
            entries[col_name] = ent
            
        def on_save():
            new_data = {col: ent.get() for col, ent in entries.items()}
            self.save_material_edits(mat_id, new_data)
            edit_win.destroy()
            
        btn_save = ttk.Button(main_frame, text="변경사항 저장", command=on_save)
        btn_save.grid(row=len(fields), column=0, columnspan=2, pady=20)

    def save_material_edits(self, mat_id, new_data):
        """Save edited material data back to the database"""
        # Update DataFrame (Robust matching)
        target_id_str = self.normalize_id(mat_id)
        mask = self.materials_df['MaterialID'].apply(self.normalize_id) == target_id_str
        idx = self.materials_df.index[mask].tolist()
        if not idx:
            messagebox.showerror("오류", "자재를 찾을 수 없습니다.")
            return
            
        for col, val in new_data.items():
            if col in self.materials_df.columns:
                # Check the actual dtype of the column
                col_dtype = self.materials_df[col].dtype
                
                # Handle different data types appropriately
                if col_dtype == 'float64' or col_dtype == 'int64':
                    # Numeric columns - convert to float, empty becomes 0.0
                    try:
                        val = float(val) if str(val).strip() else 0.0
                    except (ValueError, TypeError):
                        val = 0.0
                else:
                    # String/object columns - handle empty values
                    if val == '' or val == 'nan' or pd.isna(val):
                        val = ''
                    else:
                        val = str(val)
                
                self.materials_df.at[idx[0], col] = val
        
        # Re-check SN extraction after edit
        model = self.materials_df.at[idx[0], '모델명']
        sn = self.materials_df.at[idx[0], 'SN']
        new_model, new_sn = self.extract_sn_from_model(model, sn)
        self.materials_df.at[idx[0], '모델명'] = new_model
        self.materials_df.at[idx[0], 'SN'] = new_sn
        
        # Save to Excel
        self.save_data()
        
        # Refresh everything
        self.update_stock_view()
        self.update_registration_combos()
        self.update_material_combo()
        
        messagebox.showinfo("완료", "자재 정보가 성공적으로 수정되었습니다.")

    
    def calculate_current_stock(self, *args, **kwargs):
        from site_apps.central.src.models.material_model import calculate_current_stock_impl
        return calculate_current_stock_impl(self, *args, **kwargs)

    def update_stock_view(self, *args, **kwargs):
        from site_apps.central.src.views.stock_management_view import update_stock_view_impl
        return update_stock_view_impl(self, *args, **kwargs)

    def open_detached_stock_view(self):
        """현재 재고 현황을 별도의 팝업창(모니터링 창)으로 엽니다."""
        if hasattr(self, 'detached_windows') and 'stock' in self.detached_windows and self.detached_windows['stock']['window'].winfo_exists():
            self.detached_windows['stock']['window'].lift()
            return
            
        if not hasattr(self, 'detached_windows'):
            self.detached_windows = {}
            
        popup = tk.Toplevel(self.root)
        popup.title("📦 현재 재고 모니터링 (팝업)")
        popup.geometry("1400x800")
        
        self.detached_windows['stock'] = {'window': popup}
        
        main_frame = ttk.Frame(popup, padding=10)
        main_frame.pack(expand=True, fill='both')
        
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill='x', pady=(0, 5))
        ttk.Label(info_frame, text="💡 메인 창의 필터 및 검색어 설정이 이 팝업창에도 실시간으로 반영됩니다.", font=('Malgun Gothic', 10, 'bold'), foreground='#00529B').pack(side='left')
        ttk.Button(info_frame, text="🔄 새로고침", command=self.update_stock_view).pack(side='right')
        
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(expand=True, fill='both')
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        columns = ('No.', '회사코드', '분류', '제조사', '품목명', '모델명', 'S/N', '관리단위', '수량(현장)', '단위(현장)', '현재재고', '단가', '재고하한', '재고금액', '상태/위치', '관리품번')
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.detached_windows['stock']['tree'] = tree
        
        vsb.config(command=tree.yview)
        hsb.config(command=tree.xview)
        
        col_widths = {
            'No.': 40, '회사코드': 80, '분류': 100, '제조사': 100, '품목명': 150, 
            '모델명': 150, 'S/N': 100, '관리단위': 70, 
            '수량(현장)': 80, '단위(현장)': 70, '현재재고': 80,
            '단가': 90, '재고하한': 80, '재고금액': 120, '상태/위치': 100, '관리품번': 120
        }
        for col in columns:
            tree.heading(col, text=col, command=lambda c=col: self.treeview_sort_column(tree, c, False))
            tree.column(col, width=col_widths.get(col, 100), anchor='center', stretch=False)
            
        tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        self.enable_tree_column_drag(tree, context_menu_handler=lambda e: self._show_generic_tree_heading_context_menu(e, tree))
        
        # UI 스타일
        tree.tag_configure('deployed', background='#FFF9C4')
        tree.tag_configure('in_stock', background='')
        
        # 첫 데이터 로딩
        self.update_stock_view()
        
    def treeview_sort_column(self, tv, col, reverse):
        """Standard sorting function for Treeview columns"""
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        
        # Try to sort numerically if it looks like a number
        try:
            # Clean values like "1,234" or "EA"
            l.sort(key=lambda t: float(str(t[0]).replace(',', '').strip()) if str(t[0]).replace(',', '').strip() else 0, reverse=reverse)
        except (ValueError, TypeError):
            # Fallback to string sort
            l.sort(reverse=reverse)

        # Rearrange items in sorted positions
        for index, (val, k) in enumerate(l):
            tv.move(k, '', index)

        # Switch to reverse sort for next click
        tv.heading(col, command=lambda: self.treeview_sort_column(tv, col, not reverse))
        
    def setup_inout_tab(self):
        # Main PanedWindow (Vertical): Top (Registration) vs Bottom (History)
        self.inout_paned = ttk.Panedwindow(self.tab_inout, orient='vertical')
        self.inout_paned.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Horizontal PanedWindow for Registration Sections (Side-by-Side)
        self.inout_top_paned = ttk.Panedwindow(self.inout_paned, orient='horizontal')
        self.inout_paned.add(self.inout_top_paned, weight=1)
        
        # Save sash position on adjustment
        self.inout_paned.bind("<ButtonRelease-1>", lambda e: self.save_tab_config())
        
        # Side 1: New Material Registration
        reg_container = ttk.Frame(self.inout_top_paned)
        self.inout_top_paned.add(reg_container, weight=1)
        
        # Frame for Registration (Directly packed)
        reg_frame = ttk.LabelFrame(reg_container, text="자재 신규 등록")
        reg_frame.pack(fill='both', expand=True, padx=10, pady=5)
        reg_frame.grid_columnconfigure(1, weight=1)
        reg_frame.grid_columnconfigure(3, weight=1)
        
        # Row 0
        ttk.Label(reg_frame, text="회사코드:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.cb_co_code = ttk.Combobox(reg_frame, width=15)
        self.cb_co_code.grid(row=0, column=1, padx=5, pady=2, sticky='ew')
        self.cb_co_code.bind('<Return>', lambda e: self.cb_eq_code.focus_set())
        
        ttk.Label(reg_frame, text="설비코드:").grid(row=0, column=2, padx=5, pady=2, sticky='w')
        self.cb_eq_code = ttk.Combobox(reg_frame, width=20)
        self.cb_eq_code.grid(row=0, column=3, padx=5, pady=2, sticky='ew')
        self.cb_eq_code.bind('<Return>', lambda e: self.cb_item_name.focus_set())
        
        # Row 1
        ttk.Label(reg_frame, text="품목명:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        item_name_container = ttk.Frame(reg_frame)
        item_name_container.grid(row=1, column=1, padx=5, pady=2, sticky='ew')
        
        self.cb_item_name = ttk.Combobox(item_name_container, width=15)
        self.cb_item_name.pack(side='left', fill='x', expand=True)
        self.cb_item_name.bind('<Return>', lambda e: self.cb_model.focus_set())
        
        # [MODERN] Place the search button INSIDE the combobox
        btn_pref_search = tk.Button(self.cb_item_name, text="🔍", font=('Arial', 8), 
                                    bd=0, bg='white', cursor='hand2',
                                    command=self.open_preferred_item_search_dialog)
        # Offset x=-20 to avoid overlapping the combobox arrow
        btn_pref_search.place(relx=1.0, x=-20, rely=0.5, anchor='e', width=18, height=18)
        
        ttk.Label(reg_frame, text="모델명:").grid(row=1, column=2, padx=5, pady=2, sticky='w')
        self.cb_model = ttk.Combobox(reg_frame, width=20)
        self.cb_model.grid(row=1, column=3, padx=5, pady=2, sticky='ew')
        self.cb_model.bind('<Return>', lambda e: self.ent_sn.focus_set())
        
        # Row 2
        ttk.Label(reg_frame, text="SN번호:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.ent_sn = ttk.Entry(reg_frame, width=15)
        self.ent_sn.grid(row=2, column=1, padx=5, pady=2, sticky='ew')
        self.ent_sn.bind('<Return>', lambda e: self.cb_class.focus_set())
        
        ttk.Label(reg_frame, text="분류:").grid(row=2, column=2, padx=5, pady=2, sticky='w')
        self.cb_class = ttk.Combobox(reg_frame, width=20)
        self.cb_class.grid(row=2, column=3, padx=5, pady=2, sticky='ew')
        self.cb_class.bind('<Return>', lambda e: self.cb_spec.focus_set())
        
        # Row 3
        ttk.Label(reg_frame, text="규격:").grid(row=3, column=0, padx=5, pady=2, sticky='w')
        self.cb_spec = ttk.Combobox(reg_frame, width=15)
        self.cb_spec.grid(row=3, column=1, padx=5, pady=2, sticky='ew')
        self.cb_spec.bind('<Return>', lambda e: self.cb_unit.focus_set())
        
        ttk.Label(reg_frame, text="단위:").grid(row=3, column=2, padx=5, pady=2, sticky='w')
        self.cb_unit = ttk.Combobox(reg_frame, width=20)
        self.cb_unit.grid(row=3, column=3, padx=5, pady=2, sticky='ew')
        self.cb_unit.bind('<Return>', lambda e: self.cb_supplier.focus_set())
        
        # Row 4
        ttk.Label(reg_frame, text="공급업자:").grid(row=4, column=0, padx=5, pady=2, sticky='w')
        self.cb_supplier = ttk.Combobox(reg_frame, width=15)
        self.cb_supplier.grid(row=4, column=1, padx=5, pady=2, sticky='ew')
        self.cb_supplier.bind('<Return>', lambda e: self.cb_mfr.focus_set())
        
        ttk.Label(reg_frame, text="제조사:").grid(row=4, column=2, padx=5, pady=2, sticky='w')
        self.cb_mfr = ttk.Combobox(reg_frame, width=20)
        self.cb_mfr.grid(row=4, column=3, padx=5, pady=2, sticky='ew')
        self.cb_mfr.bind('<Return>', lambda e: self.cb_origin.focus_set())
        
        # Row 5
        ttk.Label(reg_frame, text="제조국:").grid(row=5, column=0, padx=5, pady=2, sticky='w')
        self.cb_origin = ttk.Combobox(reg_frame, width=15)
        self.cb_origin.grid(row=5, column=1, padx=5, pady=2, sticky='ew')
        self.cb_origin.bind('<Return>', lambda e: self.ent_reorder.focus_set())
        
        ttk.Label(reg_frame, text="재주문 수준:").grid(row=5, column=2, padx=5, pady=2, sticky='w')
        self.ent_reorder = ttk.Entry(reg_frame, width=20)
        self.ent_reorder.grid(row=5, column=3, padx=5, pady=2, sticky='ew')
        self.ent_reorder.insert(0, "0")
        self.ent_reorder.bind('<Return>', lambda e: self.ent_init.focus_set())
        
        # Row 6
        ttk.Label(reg_frame, text="초기재고:").grid(row=6, column=0, padx=5, pady=2, sticky='w')
        self.ent_init = ttk.Entry(reg_frame, width=15)
        self.ent_init.grid(row=6, column=1, padx=5, pady=2, sticky='ew')
        self.ent_init.insert(0, "0")
        self.ent_init.bind('<Return>', lambda e: self.ent_price.focus_set())
        
        ttk.Label(reg_frame, text="단가(가격):").grid(row=6, column=2, padx=5, pady=2, sticky='w')
        self.ent_price = ttk.Entry(reg_frame, width=20)
        self.ent_price.grid(row=6, column=3, padx=5, pady=2, sticky='ew')
        self.ent_price.insert(0, "0")
        self.ent_price.bind('<Return>', lambda e: self.ent_cost.focus_set())
        
        # Row 7
        ttk.Label(reg_frame, text="원가:").grid(row=7, column=0, padx=5, pady=2, sticky='w')
        self.ent_cost = ttk.Entry(reg_frame, width=15)
        self.ent_cost.grid(row=7, column=1, padx=5, pady=2, sticky='ew')
        self.ent_cost.insert(0, "0")
        self.ent_cost.bind('<Return>', lambda e: btn_reg.focus_set())
        
        # Row 8
        btn_reg = ttk.Button(reg_frame, text="자재 등록", command=self.register_material)
        btn_reg.grid(row=8, column=0, columnspan=4, pady=10)
        

        
        # Side 2: Transaction Entry
        trans_reg_container = ttk.Frame(self.inout_top_paned)
        self.inout_top_paned.add(trans_reg_container, weight=1)
        
        # Frame for In/Out Transaction (Directly packed)
        trans_frame = ttk.LabelFrame(trans_reg_container, text="입출고 기록")
        trans_frame.pack(fill='both', expand=True, padx=10, pady=5)
        trans_frame.grid_columnconfigure(1, weight=1)
        trans_frame.grid_columnconfigure(3, weight=1)

        # Bottom frame for history (Occupies full width at bottom)
        history_container = ttk.Frame(self.inout_paned)
        self.inout_paned.add(history_container, weight=2)
        
        ttk.Label(trans_frame, text="자재 선택:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        mat_sel_container = ttk.Frame(trans_frame)
        mat_sel_container.grid(row=0, column=1, padx=5, pady=2, columnspan=3, sticky='ew')
        
        self.cb_material = ttk.Combobox(mat_sel_container, width=45, font=('Malgun Gothic', 10))
        self.cb_material.pack(side='left', fill='x', expand=True)
        self.cb_material.bind('<<ComboboxSelected>>', self.on_material_selected)
        self.cb_material.bind('<Return>', lambda e: self.cb_type.focus_set())
        
        # [MODERN] Place the search button INSIDE the combobox
        btn_mat_search = tk.Button(self.cb_material, text="🔍", font=('Arial', 8), 
                                    bd=0, bg='white', cursor='hand2',
                                    command=self.open_material_search_dialog)
        # Offset x=-20 to avoid overlapping the combobox arrow
        btn_mat_search.place(relx=1.0, x=-20, rely=0.5, anchor='e', width=18, height=18)
        
        # Register autocomplete for material selection
        self.enable_autocomplete(self.cb_material, values_list=[]) # Initial empty, will be updated
        self.update_material_combo()
        
        # Model List display - Moved to row 4 for narrower fit
        ttk.Label(trans_frame, text="관련 모델명:").grid(row=4, column=0, padx=5, pady=2, sticky='nw')
        self.list_models = tk.Listbox(trans_frame, height=3, width=42, font=('Malgun Gothic', 9))
        self.list_models.grid(row=4, column=1, columnspan=3, padx=5, pady=2, sticky='ew')
        
        # Add scrollbar for model list
        model_vsb = ttk.Scrollbar(trans_frame, orient="vertical", command=self.list_models.yview)
        model_vsb.grid(row=4, column=4, sticky='ns', pady=2)
        self.list_models.config(yscrollcommand=model_vsb.set)
        
        ttk.Label(trans_frame, text="구분:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.cb_type = ttk.Combobox(trans_frame, values=["IN", "OUT"], state="readonly", width=15)
        self.cb_type.grid(row=1, column=1, padx=5, pady=2, sticky='ew')
        self.cb_type.set("OUT")
        self.cb_type.bind('<Return>', lambda e: self.ent_qty.focus_set())
        
        ttk.Label(trans_frame, text="수량:").grid(row=1, column=2, padx=5, pady=2, sticky='w')
        self.ent_qty = ttk.Entry(trans_frame, width=30)
        self.ent_qty.grid(row=1, column=3, padx=5, pady=2, sticky='ew')
        self.ent_qty.bind('<Return>', lambda e: self.cb_trans_site.focus_set())
        
        ttk.Label(trans_frame, text="현장:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.cb_trans_site = ttk.Combobox(trans_frame, width=28, values=self.sites)
        self.cb_trans_site.grid(row=2, column=1, padx=5, pady=2, sticky='ew')
        self.cb_trans_site.bind('<FocusOut>', lambda e: self.auto_save_to_list(e, self.cb_trans_site, self.sites, 'sites'))
        self.cb_trans_site.bind('<Return>', lambda e: self._on_trans_site_return(e))
        
        ttk.Label(trans_frame, text="창고:").grid(row=2, column=2, padx=5, pady=2, sticky='w')
        self.cb_warehouse = ttk.Combobox(trans_frame, width=28, values=self.warehouses)
        self.cb_warehouse.grid(row=2, column=3, padx=5, pady=2, sticky='ew')
        self.cb_warehouse.bind('<FocusOut>', lambda e: self.auto_save_to_list(e, self.cb_warehouse, self.warehouses, 'warehouses'))
        self.cb_warehouse.bind('<Return>', lambda e: self._on_warehouse_return(e))
        
        ttk.Label(trans_frame, text="담당자:").grid(row=3, column=0, padx=5, pady=2, sticky='w')
        self.ent_user = ttk.Combobox(trans_frame, width=28, values=getattr(self, 'users', []))
        self.ent_user.grid(row=3, column=1, padx=5, pady=2, sticky='ew')
        self.ent_user.bind('<FocusOut>', lambda e: self.auto_save_to_list(e, self.ent_user, self.users, 'users'))
        self.ent_user.bind('<Return>', lambda e: self._on_user_return(e))
        
        ttk.Label(trans_frame, text="비고:").grid(row=3, column=2, padx=5, pady=2, sticky='w')
        self.ent_note = ttk.Entry(trans_frame, width=30)
        self.ent_note.grid(row=3, column=3, padx=5, pady=2, sticky='ew')
        self.ent_note.bind('<Return>', lambda e: btn_trans.focus_set())
        
        btn_trans = ttk.Button(trans_frame, text="기록 저장", command=self.add_transaction)
        btn_trans.grid(row=5, column=0, columnspan=4, pady=10)
        
        # Frame for displaying transaction history
        history_frame = ttk.LabelFrame(history_container, text="최근 입출고 내역")
        history_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Filter/Control frame for history
        history_ctrl_frame = ttk.Frame(history_frame)
        history_ctrl_frame.pack(fill='x', padx=5, pady=2)
        
        # First row of filters
        filter_row1 = ttk.Frame(history_ctrl_frame)
        filter_row1.pack(fill='x', pady=2)
        
        btn_del_trans = ttk.Button(filter_row1, text="선택 항목 삭제", command=self.delete_transaction_entry)
        btn_del_trans.pack(side='left', padx=5)
        
        btn_export_trans = ttk.Button(filter_row1, text="엑셀 내보내기", command=self.export_transaction_history)
        btn_export_trans.pack(side='left', padx=5)

        btn_refresh_trans = ttk.Button(filter_row1, text="입출고 내역 새로고침", command=self.refresh_inout_history)
        btn_refresh_trans.pack(side='left', padx=5)

        # [NEW] Cleanup button for automatic transactions (Equipment only)
        btn_cleanup_trans = ttk.Button(filter_row1, text="🧹 장비류 자동기록 클린업", command=self.cleanup_auto_transactions)
        btn_cleanup_trans.pack(side='left', padx=10)
        
        # Second row of filters
        filter_row2 = ttk.Frame(history_ctrl_frame)
        filter_row2.pack(fill='x', pady=2)
        
        ttk.Label(filter_row2, text="품목명 필터:").pack(side='left', padx=5)
        self.cb_trans_filter_mat = ttk.Combobox(filter_row2, width=35, state="readonly")
        self.cb_trans_filter_mat.pack(side='left', padx=5)
        self.cb_trans_filter_mat.bind('<<ComboboxSelected>>', lambda e: self.update_transaction_view())
        
        ttk.Label(filter_row2, text="현장 필터:").pack(side='left', padx=(20, 5))
        self.cb_trans_filter_site = ttk.Combobox(filter_row2, width=12, state="readonly")
        self.cb_trans_filter_site.pack(side='left', padx=5)
        self.cb_trans_filter_site.bind('<<ComboboxSelected>>', lambda e: self.update_transaction_view())
        
        # Third row of filters
        filter_row3 = ttk.Frame(history_ctrl_frame)
        filter_row3.pack(fill='x', pady=2)
        
        ttk.Label(filter_row3, text="차량번호 필터:").pack(side='left', padx=5)
        self.cb_trans_filter_vehicle = ttk.Combobox(filter_row3, width=12, state="readonly")
        self.cb_trans_filter_vehicle.pack(side='left', padx=5)
        self.cb_trans_filter_vehicle.bind('<<ComboboxSelected>>', lambda e: self.update_transaction_view())
        self.cb_trans_filter_vehicle.bind('<F5>', lambda e: (self.refresh_inout_history(), 'break')[1])
        
        # Treeview for history
        tree_scroll_frame = ttk.Frame(history_frame)
        tree_scroll_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        inout_vsb = ttk.Scrollbar(tree_scroll_frame, orient="vertical")
        inout_hsb = ttk.Scrollbar(tree_scroll_frame, orient="horizontal")
        
        # [RESTRUCTURED] Optimized for NDT consumables: Focus on RT, PT, MT quantities
        columns = ('날짜', '현장', '구분', '품목명', '수량', '세척제', '침투제', '현상제', '백색페인트', '흑색자분', 'SN', '규격', '창고', '담당자', '비고', '검사비')
        self.inout_tree = ttk.Treeview(tree_scroll_frame, columns=columns, show='headings', height=10,
                                       yscrollcommand=inout_vsb.set, xscrollcommand=inout_hsb.set)
        
        inout_vsb.config(command=self.inout_tree.yview)
        inout_hsb.config(command=self.inout_tree.xview)
        
        # [NEW] Default Column Widths for NDT-focused View
        col_widths = {
            '날짜': 130, '현장': 120, '구분': 60, '품목명': 160, 
            '수량': 70, '세척제': 60, '침투제': 60, '현상제': 60, '백색페인트': 80, '흑색자분': 80,
            'SN': 100, '규격': 100, '창고': 80, '담당자': 80, '비고': 150, '검사비': 90
        }
        for col in columns:
            self.inout_tree.heading(col, text=col)
            width = col_widths.get(col, 100)
            self.inout_tree.column(col, width=width, minwidth=50, stretch=False, anchor='center')
        
        # [FIX] Select only the core NDT-relevant columns for initial display
        visible_cols = ('날짜', '현장', '구분', '품목명', '수량', '세척제', '침투제', '현상제', '백색페인트', '흑색자분', '비고')
        self.inout_tree['displaycolumns'] = visible_cols
        self.enable_tree_column_drag(self.inout_tree, context_menu_handler=lambda e: self._show_generic_tree_heading_context_menu(e, self.inout_tree))
        self.inout_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config(), add='+')
        self.inout_tree.bind('<F5>', lambda e: (self.refresh_inout_history(), 'break')[1])

        self.inout_tree.grid(row=0, column=0, sticky='nsew')
        inout_vsb.grid(row=0, column=1, sticky='ns')
        inout_hsb.grid(row=1, column=0, sticky='ew')
        
        tree_scroll_frame.grid_rowconfigure(0, weight=1)
        tree_scroll_frame.grid_columnconfigure(0, weight=1)
        
        # Initial populate
        self.update_transaction_view()
        
        # Set initial sash position for better balance
        self.inout_paned.after(200, self._ensure_inout_sash_visibility)

    def _ensure_inout_sash_visibility(self):
        """Ensure the inout sash position is properly set"""
        try:
            if hasattr(self, 'inout_paned'):
                self.inout_paned.update_idletasks()
                total_h = self.inout_paned.winfo_height()
                if total_h > 200:
                    # Set vertical sash to 40% of total height (registration area smaller)
                    new_pos = int(total_h * 0.4)
                    self.inout_paned.sashpos(0, new_pos)
                    
                    # Set horizontal sash for side-by-side
                    total_w = self.inout_top_paned.winfo_width()
                    if total_w > 200:
                        self.inout_top_paned.sashpos(0, int(total_w * 0.45))
                    print(f"Set inout sash to {new_pos} (total height: {total_h})")
        except Exception as e:
            print(f"Error ensuring inout sash visibility: {e}")


    def get_material_display_name(self, mat_id):
        """Get formatted material name as '품목명 (SN: SN번호) - 규격'"""
        # [FIX] Handle NaN IDs gracefully to prevent "NAN" display
        if pd.isna(mat_id) or str(mat_id).lower().strip() == 'nan':
            return ""
            
        if self.materials_df.empty:
            return f"ID: {mat_id}"
            
        try:
            mat_row = self.materials_df[self.materials_df['MaterialID'] == mat_id]
        except (TypeError, ValueError):
            mat_row = pd.DataFrame()  # Comparison failed (e.g. dtype mismatch) → treat as not found
            
        if mat_row.empty:
            # [NEW] Handle non-numeric IDs (raw names for PAUT) and orphans gracefully
            return str(mat_id)
            
        mat = mat_row.iloc[0]
        name = mat['품목명']
        
        # [SAFETY] If 품목명 is NaN/empty, fall back to str(mat_id)
        if pd.isna(name) or str(name).strip() == '':
            return str(mat_id)
        
        sn = mat.get('SN', '')
        spec = mat.get('규격', '')
        
        # Build display string: 품목명-SN (모델명) [규격]
        display = name
        if sn and pd.notna(sn) and str(sn).strip():
            display = f"{name}-{str(sn).strip()}"
            
        model_name = mat.get('모델명', '')
        model_name_val = str(model_name).strip()
        if model_name_val and pd.notna(model_name) and model_name_val != str(name).strip():
            display = f"{display} ({model_name_val})"
        
        if spec and pd.notna(spec) and str(spec).strip():
            display = f"{display} [{str(spec).strip()}]"
            
        return display


    def get_material_name_only(self, mat_id):
        """Get only the 품목명 for a material"""
        if self.materials_df.empty:
            return f"(미등록) {mat_id}"
        
        target_id = self.normalize_id(mat_id)
        # Match using normalized IDs to handle formatting variations
        mask = self.materials_df['MaterialID'].apply(self.normalize_id) == target_id
        mat_row = self.materials_df[mask]
        
        if mat_row.empty:
            return f"(미등록) {mat_id}"
        return str(mat_row.iloc[0].get('품목명', f"(미등록) {mat_id}"))

    def get_material_info(self, mat_id):
        """Get all material columns as a dictionary"""
        if self.materials_df.empty:
            return {}
        
        target_id = self.normalize_id(mat_id)
        mask = self.materials_df['MaterialID'].apply(self.normalize_id) == target_id
        mat_row = self.materials_df[mask]
        
        if mat_row.empty:
            return {}
        return mat_row.iloc[0].to_dict()
        
        # Add specification if exists
        if spec and pd.notna(spec) and str(spec).strip():
            display += f" - {spec}"
            
        return display

    def update_material_combo(self):
        """Update both In/Out and Daily Usage material comboboxes with unified list"""
        mat_list = []
        if not self.materials_df.empty:
            # Create list with unified display format
            # [STRICT] Robust check for Active status
            active_mask = pd.to_numeric(self.materials_df.get('Active', 1), errors='coerce').fillna(1) > 0.5
            active_df = self.materials_df[active_mask]
            
            for _, mat in active_df.iterrows():
                display = self.get_material_display_name(mat['MaterialID'])
                mat_list.append(display)
        
        # [NEW] Pre-calculate inactive items for efficient filtering (STRICT)
        inactive_ids = set()
        inactive_names = set()
        if not self.materials_df.empty:
            # Use strict numeric check
            active_vals = pd.to_numeric(self.materials_df.get('Active', 1), errors='coerce').fillna(1)
            for idx, m in self.materials_df.iterrows():
                if active_vals.loc[idx] < 0.5:
                    inactive_ids.add(self.normalize_id(m.get('MaterialID')))
                    inactive_names.add(str(m.get('품목명', '')).strip())

        # [NEW] Also collect historical material names from daily usage history (Only if still Active)
        history_materials = []
        if not self.daily_usage_df.empty and 'MaterialID' in self.daily_usage_df.columns:
            unique_mat_ids = self.daily_usage_df['MaterialID'].dropna().unique()
            for mat_id in unique_mat_ids:
                str_id = self.normalize_id(mat_id)
                if str_id in inactive_ids:
                    continue
                name = self.get_material_display_name(mat_id)
                if name: history_materials.append(name)
        
        # Merge database items with centralized films list and historical materials, unique and sort
        all_vals = list(set([str(m) for m in mat_list + self.carestream_films + history_materials if pd.notna(m) and str(m).strip()]))
        all_vals.sort()
        
        # Update ComboBoxes
        self.cached_material_list = all_vals
        if hasattr(self, 'cb_material'):
            self.cb_material['values'] = all_vals
        if hasattr(self, 'cb_daily_material'):
            if isinstance(self.cb_daily_material, ttk.Combobox):
                self.cb_daily_material['values'] = all_vals
                try:
                    max_chars = max((len(str(v)) for v in all_vals), default=12)
                    self.cb_daily_material.configure(width=min(40, max(22, max_chars + 2)))
                except Exception:
                    pass
            
        if hasattr(self, 'cb_daily_equip'):
            # [NEW] Collect history of equipment names from daily usage records (Only if Active)
            history_equip = []
            if not self.daily_usage_df.empty and '장비명' in self.daily_usage_df.columns:
                raw_history = self.daily_usage_df['장비명'].dropna().unique().tolist()
                for eq in raw_history:
                    eq_str = str(eq).strip()
                    if not eq_str or eq_str in inactive_names:
                        continue
                    history_equip.append(eq_str)
                
            # [NEW] Filter: Only include actual equipment (non-consumables) in the equipment dropdown
            equip_only_vals = []
            for v in all_vals:
                base_name = v.split('[')[0].split('-')[0].strip()
                if not self._is_consumable_material(base_name, ''):
                    equip_only_vals.append(v)
                
            # Combine custom equipments, history, and the filtered equipment list
            combined_equip = list(set([str(e).strip() for e in self.equipments + history_equip + equip_only_vals if pd.notna(e) and str(e).strip()]))
            combined_equip.sort()
            
            # [FIX] Only set values if it's a Combobox
            if isinstance(self.cb_daily_equip, ttk.Combobox):
                self.cb_daily_equip['values'] = combined_equip
            
            try:
                max_chars = max((len(str(v)) for v in combined_equip), default=12)
                self.cb_daily_equip.configure(width=min(40, max(22, max_chars + 2)))
            except Exception:
                pass
            self.equipment_suggestions = combined_equip
        
        self.materials_display_list = all_vals
        
        # [NEW] Calculate consumable-only list for Inventory Status filters (Exclude PAUT, etc.)
        self.consumable_display_list = []
        for val in all_vals:
            # Extract the base name (before spec '-' or ID '[')
            base_name = val.split('[')[0].split('-')[0].strip()
            if self._is_consumable_material(base_name, ''):
                self.consumable_display_list.append(val)
        self.consumable_display_list.sort()


        if hasattr(self, 'cb_trans_filter_mat'):
            self.cb_trans_filter_mat['values'] = ["전체"] + all_vals
            if not self.cb_trans_filter_mat.get():
                self.cb_trans_filter_mat.set("전체")
        
        if hasattr(self, 'cb_trans_filter_site'):
            self.cb_trans_filter_site['values'] = ["전체"] + sorted(self.sites)
            if not self.cb_trans_filter_site.get():
                self.cb_trans_filter_site.set("전체")
        
        if hasattr(self, 'cb_trans_filter_vehicle'):
            # Get unique vehicle numbers from transactions_df
            if '차량번호' in self.transactions_df.columns:
                vehicle_numbers = self.transactions_df['차량번호'].dropna().astype(str).str.strip().unique()
                vehicle_list = sorted([v for v in vehicle_numbers if v and v != 'nan'])
            else:
                vehicle_list = []
            
            # Add sample vehicle numbers for testing if no data exists
            if not vehicle_list:
                vehicle_list = ['12가1234', '34나5678', '89다9012']
            
            self.cb_trans_filter_vehicle['values'] = ["전체"] + vehicle_list
            if not self.cb_trans_filter_vehicle.get():
                self.cb_trans_filter_vehicle.set("전체")

    def _get_material_candidates(self, include_all=False):
        vals = []
        if hasattr(self, 'materials_display_list'):
            vals.extend(getattr(self, 'materials_display_list', []) or [])
        if hasattr(self, 'cb_daily_material'):
            pass # Removed reading from widget to prevent recursive filtering issues

        cleaned = sorted(set(str(v).strip() for v in vals if str(v).strip() and str(v).strip() != '전체' and not str(v).strip().startswith('ID: ')))
        return (['전체'] + cleaned) if include_all else cleaned

    def _get_history_material_candidates(self, include_all=False):
        """Get material names from both current inventory and history for selection"""
        raw_materials = set()
        
        # 1. Add all active materials from stock master (Primary Source)
        if not self.materials_df.empty:
            for _, mat in self.materials_df.iterrows():
                if mat.get('Active', 1) == 1:
                    name = self.get_material_display_name(mat['MaterialID'])
                    if name: raw_materials.add(name)
        
        # 2. Add materials exclusively from daily usage history (Secondary Source)
        if not self.daily_usage_df.empty and 'MaterialID' in self.daily_usage_df.columns:
            unique_mat_ids = self.daily_usage_df['MaterialID'].dropna().unique()
            for mat_id in unique_mat_ids:
                name = self.get_material_display_name(mat_id)
                if name: raw_materials.add(name)
        
        cleaned = sorted(list(raw_materials))
        return (['전체'] + cleaned) if include_all else cleaned

    def _get_equipment_candidates(self, include_all=False):
        """Get unified equipment suggestions from master list and history"""
        raw_equip = set()
        
        # 1. Base equipment list
        if hasattr(self, 'equipments'):
            for e in getattr(self, 'equipments', []):
                if e and str(e).strip(): raw_equip.add(str(e).strip())
                
        # 2. Dynamic suggestions
        if hasattr(self, 'equipment_suggestions'):
            for e in getattr(self, 'equipment_suggestions', []):
                if e and str(e).strip(): raw_equip.add(str(e).strip())
                
        # 3. History
        if hasattr(self, 'daily_usage_df') and not self.daily_usage_df.empty and '장비명' in self.daily_usage_df.columns:
            unique_equip = self.daily_usage_df['장비명'].dropna().astype(str).str.strip().unique()
            for e in unique_equip:
                if e: raw_equip.add(e)

        cleaned = sorted([e for e in raw_equip if e != '전체'])
        return (['전체'] + cleaned) if include_all else cleaned

    def _get_inspection_item_candidates(self):
        """Get candidates for Inspection Item (Inspection Item Name)"""
        items = set(getattr(self, 'test_items', []))
        if not self.daily_usage_df.empty and '검사품명' in self.daily_usage_df.columns:
            hist = self.daily_usage_df['검사품명'].dropna().astype(str).str.strip().unique()
            for h in hist: 
                if h: items.add(h)
        return sorted(list(items))

    def _get_applied_code_candidates(self):
        """Get candidates for Applied Code"""
        codes = set(getattr(self, 'applied_codes', []))
        if not self.daily_usage_df.empty and '적용코드' in self.daily_usage_df.columns:
            hist = self.daily_usage_df['적용코드'].dropna().astype(str).str.strip().unique()
            for h in hist: 
                if h: codes.add(h)
        return sorted(list(codes))

    def _apply_combobox_word_suggest(self, combobox, source_values, open_dropdown=False):
        if not combobox or not isinstance(combobox, ttk.Combobox):
            return

        text_raw = combobox.get()
        text = text_raw.strip()
        had_focus = False
        try:
            # More robust focus check
            curr_focus = self.root.focus_get()
            had_focus = (curr_focus == combobox or str(curr_focus).startswith(str(combobox)))
        except Exception:
            had_focus = False
        
        try:
            cursor_pos = combobox.index(tk.INSERT)
        except Exception:
            cursor_pos = len(text_raw)
            
        source = [str(v).strip() for v in (source_values or []) if str(v).strip()]

        if not text:
            filtered = source
        else:
            q = text.lower()
            filtered = [v for v in source if q in v.lower()]

        # [FIX] Only update if values changed to avoid recursion/flicker
        values_changed = list(combobox['values']) != list(filtered)
        if values_changed:
            combobox['values'] = filtered

        # Always ensure state is reconciled if it had focus
        if had_focus:
            def _restore_state():
                try:
                    if self.root.focus_get() != combobox:
                        combobox.focus_set()
                    combobox.set(text_raw)
                    combobox.icursor(cursor_pos)
                    combobox.selection_clear()
                except Exception:
                    pass

            _restore_state()
            # Double tap with after_idle to ensure it sticks after the internal Tk events
            try: combobox.after_idle(_restore_state)
            except Exception: pass
        
        # [REVISION] Use 'Alt-Down' for more stable dropdown posting
        if open_dropdown and had_focus and filtered and len(text) > 0:
            try:
                # [STABILITY] Don't re-open if just selected
                if getattr(combobox, '_just_selected', False):
                    return
                # Alt-Down is the standard shortcut to post the dropdown in many OS/Toolkits
                combobox._auto_opening = True
                combobox.event_generate('<Alt-Down>')
                combobox._auto_opening = False
                # Re-verify cursor immediately after posting
                combobox.icursor(cursor_pos)
            except:
                pass

    def _bind_combobox_word_suggest(self, combobox, source_getter):
        if not combobox:
            return

        def _on_interaction(e=None, open=True):
            # Skip for navigation keys which have native behaviors
            if e is not None and hasattr(e, 'keysym'):
                if e.keysym in {
                    'Left', 'Right', 'Tab', 'Up', 'Down', 'Return', 'Escape',
                    'Shift_L', 'Shift_R', 'Control_L', 'Control_R',
                    'Alt_L', 'Alt_R', 'Caps_Lock'
                }:
                    return
                # [NEW] Skip 'Process' keysym used by Korean IME during composition
                if e.keysym == 'Process':
                    return
            
            # [FIX] Debounce: Wait 300ms before showing suggestions to avoid interrupting continuous typing (especially for IME)
            if hasattr(combobox, '_suggest_after_id'):
                combobox.after_cancel(combobox._suggest_after_id)
            
            def _do_suggest():
                if combobox.winfo_exists():
                    self._apply_combobox_word_suggest(combobox, source_getter(), open_dropdown=open)
            
            combobox._suggest_after_id = self.root.after(300, _do_suggest)

        combobox.bind('<KeyRelease>', lambda e: _on_interaction(e, open=True), add='+')
        combobox.bind('<FocusIn>', lambda e: _on_interaction(e, open=False), add='+')
        
        # [NEW] Guard against "sticky" dropdown after selection
        def _on_selected(e=None):
            combobox._just_selected = True
            def _reset(): combobox._just_selected = False
            combobox.after(500, _reset)
            
        combobox.bind('<<ComboboxSelected>>', _on_selected, add='+')
        
        # [NEW] Restore full list when user clicks anywhere on the combobox (including dropdown arrow)
        def _on_click(e=None):
            if not getattr(combobox, '_auto_opening', False):
                try:
                    combobox['values'] = source_getter()
                except: pass
        combobox.bind('<ButtonPress-1>', _on_click, add='+')

    def update_registration_combos(self):
        """Update registration comboboxes with unique values from database and centralized list"""
        # 1. Update registration fields from database
        fields = {}
        for key, attr in [
            ('회사코드', 'cb_co_code'),
            ('관리품번', 'cb_eq_code'),
            ('품목명', 'cb_item_name'),
            ('품목군코드', 'cb_class'),
            ('규격', 'cb_spec'),
            ('관리단위', 'cb_unit'),
            ('공급업체', 'cb_supplier'),
            ('제조사', 'cb_mfr'),
            ('제조국', 'cb_origin')
        ]:
            if hasattr(self, attr):
                fields[key] = getattr(self, attr)
        
        # Store lists for autocomplete (Main Catalog + History)
        self.co_code_list = []
        self.eq_code_list = []
        self.item_name_list = []
        self.class_list = []
        self.spec_list = []
        self.unit_list = []
        self.supplier_list = []
        self.mfr_list = []
        self.origin_list = []
        self.sn_list = []
        self.model_list = []

        attr_mapping = {
            '회사코드': 'co_code_list',
            '관리품번': 'eq_code_list',
            '품목명': 'item_name_list',
            '품목군코드': 'class_list',
            '규격': 'spec_list',
            '관리단위': 'unit_list',
            '공급업체': 'supplier_list',
            '제조사': 'mfr_list',
            '제조국': 'origin_list',
            'SN': 'sn_list',
            '모델명': 'model_list'
        }

        for col, list_attr in attr_mapping.items():
            vals = []
            if not self.materials_df.empty and col in self.materials_df.columns:
                # [STRICT] Robust check for Active status (numeric comparison)
                # Ensure Active is treated as numeric, defaulting to 1 (Active)
                active_mask = pd.to_numeric(self.materials_df.get('Active', 1), errors='coerce').fillna(1) > 0.5
                active_df = self.materials_df[active_mask]
                unique_vals = active_df[col].dropna().unique()
                vals = sorted([str(v).strip() for v in unique_vals if v and str(v).strip()])
            
            # Update instance attributes for autocomplete
            if col == '관리단위':
                # Merge DB units with the managed daily_units
                self.unit_list = list(dict.fromkeys(self.daily_units + vals))
                if hasattr(self, 'cb_daily_unit'):
                    self.cb_daily_unit['values'] = self.unit_list
            else:
                setattr(self, list_attr, vals)
            
            # 1. Update registration fields (without "전체")
            if col in fields:
                fields[col]['values'] = vals

            # 2. Update Stock View filters (with "전체")
            # [NEW] For Stock View filters, also exclude non-consumable equipment to match the list view
            filter_map = {
                '회사코드': getattr(self, 'cb_filter_co', None),
                '품목군코드': getattr(self, 'cb_filter_class', None),
                '제조사': getattr(self, 'cb_filter_mfr', None),
                '품목명': getattr(self, 'cb_filter_name', None),
                'SN': getattr(self, 'cb_filter_sn', None),
                '모델명': getattr(self, 'cb_filter_model', None),
                '관리품번': getattr(self, 'cb_filter_eq', None)
            }
            if col in filter_map and filter_map[col] is not None:
                combo = filter_map[col]
                
                # Filter vals for consumables if it's the item name or model
                c_vals = vals
                if col in ['품목명', '모델명', '관리품번']:
                    c_vals = []
                    for v in vals:
                        # Heuristic: Check if any item with this name/model is a consumable
                        if col == '품목명':
                            if self._is_consumable_material(v, ''): c_vals.append(v)
                        elif col == '모델명':
                            # Check if at least one material with this model is a consumable
                            m_rows = self.materials_df[self.materials_df['모델명'].astype(str).str.strip() == v]
                            if not m_rows.empty and self._is_consumable_material(m_rows.iloc[0].get('품목명', ''), ''):
                                c_vals.append(v)
                        else:
                            c_vals.append(v)
                
                combo['values'] = ["전체"] + c_vals
                if not combo.get():
                    combo.set("전체")
        
        # Add Carestream film options to 품목명 (Registration tab only)
        if hasattr(self, 'cb_item_name'):
            all_mat_vals = sorted(list(set([str(mat) for mat in self.item_name_list + self.carestream_films if pd.notna(mat) and str(mat).strip()])))
            self.item_name_list[:] = all_mat_vals
            self.cb_item_name['values'] = self.item_name_list
            
        # [FIX] Do NOT overwrite cb_filter_name values here; it was already set with consumable filtering above.
        # Just ensure the autocomplete list is refreshed.
        if hasattr(self, 'apply_autocomplete_to_all_comboboxes'):
            self.apply_autocomplete_to_all_comboboxes()
        
        # 3. Final view update now that filters are set to "전체"
        import threading
        self.root.after(100, self.update_stock_view)

    def register_material(self):
        co_code = self.cb_co_code.get()
        item_name = self.cb_item_name.get()
        eq_code = self.cb_eq_code.get()
        sn = self.ent_sn.get()
        classification = self.cb_class.get()
        spec = self.cb_spec.get()
        unit = self.cb_unit.get()
        supplier = self.cb_supplier.get()
        manufacturer = self.cb_mfr.get()
        origin = self.cb_origin.get()
        
        try:
            init_stock = float(self.ent_init.get())
            reorder_point = float(self.ent_reorder.get())
            price_val = float(self.ent_price.get() if self.ent_price.get() else 0)
            cost_val = float(self.ent_cost.get() if self.ent_cost.get() else 0)
        except ValueError:
            messagebox.showwarning("입력 오류", "재고, 단가, 원가 정보는 숫자여야 합니다.")
            return
            
        if not item_name:
            messagebox.showwarning("입력 오류", "품목명을 입력해주세요.")
            return
        
        # Generate MaterialID
        if self.materials_df.empty:
            mat_id = 1
        else:
            mat_id = self.materials_df['MaterialID'].max() + 1
        
        # Extract SN from Model Name (Now using newly added model name combobox)
        model_name = self.cb_model.get().strip() if hasattr(self, 'cb_model') else ''
        new_model, new_sn = self.extract_sn_from_model(model_name, sn)
        # Note: In register_material, model_name is currently not an input field, but SN is.
        # If the user adds model name to registration later, this will handle it.
        
        new_row = {
            'MaterialID': mat_id,
            '회사코드': co_code,
            '관리품번': eq_code,
            '품목명': item_name,
            'SN': new_sn,
            '창고': '',
            '모델명': new_model,
            '규격': spec,
            '품목군코드': classification,
            '공급업체': supplier,
            '제조사': manufacturer,
            '제조국': origin,
            '가격': price_val,
            '원가': cost_val,
            '관리단위': unit if unit else 'EA',
            '수량': init_stock,
            '재고하한': reorder_point
        }
        
        self.materials_df = pd.concat([self.materials_df, pd.DataFrame([new_row])], ignore_index=True)
        self.save_data()
        self.update_material_combo()
        self.update_registration_combos()
        self.update_stock_view()
        messagebox.showinfo("완료", f"'{item_name}' 자재가 등록되었습니다.")
        
        # Clear entries
        self.cb_co_code.set('')
        self.cb_eq_code.set('')
        self.cb_item_name.set('')
        self.ent_sn.delete(0, tk.END)
        self.cb_class.set('')
        self.cb_spec.set('')
        self.cb_unit.set('')
        self.cb_supplier.set('')
        self.cb_mfr.set('')
        self.cb_origin.set('')
        self.ent_reorder.delete(0, tk.END)
        self.ent_reorder.insert(0, "0")
        self.ent_init.delete(0, tk.END)
        self.ent_init.insert(0, "0")
        # Clear additional fields
        if hasattr(self, 'cb_model'):
            self.cb_model.set('')
        if hasattr(self, 'ent_price'):
            self.ent_price.delete(0, tk.END)
        if hasattr(self, 'ent_cost'):
            self.ent_cost.delete(0, tk.END)

    def add_transaction(self):
        """Record an IN or OUT transaction"""
        try:
            mat_selection = self.cb_material.get().strip()
            t_type = self.cb_type.get()
            user = self.ent_user.get().strip()
            
            if not mat_selection:
                messagebox.showwarning("입력 오류", "자재를 선택해주세요.")
                return
            if not t_type:
                messagebox.showwarning("입력 오류", "구분(입고/출고)을 선택해주세요.")
                return
                
            try:
                qty_str = self.ent_qty.get().strip()
                if not qty_str:
                    messagebox.showwarning("입력 오류", "수량을 입력해주세요.")
                    return
                qty = float(qty_str)
            except ValueError:
                messagebox.showwarning("입력 오류", "수량은 숫자여야 합니다.")
                return
                
            note = self.ent_note.get().strip()
            
            # Extract pure material name from selection
            mat_name = mat_selection
            if " - " in mat_name: mat_name = mat_name.split(" - ")[0]
            if " (SN: " in mat_name: mat_name = mat_name.split(" (SN: ")[0]
            pure_mat_name = mat_name.strip()
            
            # Find MaterialID
            # Ensure MaterialID is treated consistently
            mat_rows = self.materials_df[self.materials_df['품목명'] == pure_mat_name]
            
            if mat_rows.empty:
                # If exact match fails, try a case-insensitive or stripped match
                mat_rows = self.materials_df[self.materials_df['품목명'].str.strip() == pure_mat_name]
            
            if mat_rows.empty:
                # [NEW] Ask to register new material instead of just showing error
                confirm = messagebox.askyesno("신규 자재", f"'{pure_mat_name}'은(는) 등록되지 않은 자재입니다. 신규 자재로 등록하고 진행할까요?")
                if confirm:
                    mat_id = self.register_new_material(pure_mat_name)
                else:
                    return
            else:
                mat_id = mat_rows['MaterialID'].values[0]
            
            # Update Warehouse in materials_df
            warehouse = str(self.cb_warehouse.get()).strip()
            if warehouse:
                # Type-safe assignment
                if '창고' in self.materials_df.columns:
                    mask = self.materials_df['MaterialID'] == mat_id
                    if mask.any():
                        self.materials_df.loc[mask, '창고'] = warehouse
            
            # Create transaction record
            new_trans = {
                'Date': datetime.datetime.now(),
                'MaterialID': mat_id,
                'Type': t_type,
                'Quantity': qty if t_type == 'IN' else -qty,
                'Note': note,
                'User': user,
                'Site': self.cb_trans_site.get() if hasattr(self, 'cb_trans_site') else '',
                '차량번호': '',
                '주행거리': '',
                '차량점검': '',
                '차량비고': ''
            }
            
            # Add to dataframe
            self.transactions_df = pd.concat([self.transactions_df, pd.DataFrame([new_trans])], ignore_index=True)
            
            # Save all data (force sync)
            self.save_data()
            
            # Check if it was really added (for debug feedback)
            last_count = len(self.transactions_df)
            
            # Refresh views
            self.update_stock_view()
            self.update_transaction_view()
            self.update_material_combo() 
            
            # Auto-save site/user to lists
            site_value = new_trans['Site']
            if site_value and site_value not in self.sites:
                self.sites.append(site_value)
                self.sites.sort()
                self.save_tab_config()
            
            if user and user not in self.users:
                self.users.append(user)
                self.users.sort()
                self.save_tab_config()
            
            # Success feedback
            messagebox.showinfo("완료", f"{pure_mat_name} {t_type} 처리되었습니다.\n(전체 기록 수: {last_count}개)")
            
            # Clear UI fields
            self.ent_qty.delete(0, tk.END)
            self.ent_note.delete(0, tk.END)
            if hasattr(self, 'ent_user') and hasattr(self.ent_user, 'set'):
                self.ent_user.set('')
            elif hasattr(self, 'ent_user'):
                self.ent_user.delete(0, tk.END)
            
            if hasattr(self, 'cb_warehouse'): self.cb_warehouse.set('')
            if hasattr(self, 'cb_trans_site'): self.cb_trans_site.set('')
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            self.show_error_dialog("저장 오류", f"기록 저장 중 기술적인 오류가 발생했습니다:\n{e}\n\n상세 정보:\n{error_details}")
        
        # Ensure view is updated regardless
        self.update_transaction_view()

    def update_transaction_view(self):
        """Unified transaction view that combines manual stock transactions and site usage."""
        try:
            # 0. Helper Functions (Global normalize_id is now used)
            def _f(val):
                if val is None or pd.isna(val): return 0.0
                try: 
                    s = str(val).replace(',', '').strip()
                    return float(s) if s else 0.0
                except: return 0.0
                
            def _q_str(val):
                try:
                    v = abs(float(str(val).replace(',', '').strip()))
                    return f"{v:g}" if v > 0 else ""
                except: return ""

            # 1. Clear existing view
            for item in self.inout_tree.get_children():
                self.inout_tree.delete(item)
            
            # 2. Prepare Transaction Records
            df_trans = pd.DataFrame()
            active_ids = []
            if not self.materials_df.empty and 'MaterialID' in self.materials_df.columns:
                # Get list of MaterialIDs that are NOT deleted (Active != 0)
                active_mask = (self.materials_df.get('Active', 1) != 0)
                active_ids = [self.normalize_id(x) for x in self.materials_df[active_mask]['MaterialID'].dropna().tolist()]
                
            if not self.transactions_df.empty:
                df_trans = self.transactions_df.copy()
                # Column Compatibility
                if 'Site' not in df_trans.columns and '현장' in df_trans.columns: df_trans['Site'] = df_trans['현장']
                if 'Date' not in df_trans.columns and '날짜' in df_trans.columns: df_trans['Date'] = df_trans['날짜']
                if 'Date' not in df_trans.columns and '일자' in df_trans.columns: df_trans['Date'] = df_trans['일자']
                
                df_trans['Date'] = pd.to_datetime(df_trans['Date'], errors='coerce')
                if 'EntryDate' in df_trans.columns:
                    df_trans['Date'] = df_trans['Date'].fillna(pd.to_datetime(df_trans['EntryDate'], errors='coerce'))
                
                if active_ids:
                    df_trans = df_trans[df_trans['MaterialID'].apply(self.normalize_id).isin(active_ids)]
            
            # 2b. Prepare Daily Usage Records
            df_daily = pd.DataFrame()
            if hasattr(self, 'daily_usage_df') and not self.daily_usage_df.empty:
                df_daily = self.daily_usage_df.copy()
                # [FIX] Normalize column names to ensure reliable extraction of NDT_ fields
                df_daily.columns = [str(c).strip().replace(' ', '') for c in df_daily.columns]
                
                df_daily['Date'] = pd.to_datetime(df_daily['Date'], errors='coerce')
                # If Date is missing but EntryTime exists, use EntryTime
                if 'EntryTime' in df_daily.columns:
                    df_daily['Date'] = df_daily['Date'].fillna(pd.to_datetime(df_daily['EntryTime'], errors='coerce'))
                
            print(f"DEBUG: update_transaction_view - TRANS: {len(df_trans)}, DAILY: {len(df_daily)}")
            def _f(val):
                if val is None or pd.isna(val): return 0.0
                try: 
                    s = str(val).replace(',', '').strip()
                    return float(s) if s else 0.0
                except: return 0.0
            # [FIX] Do NOT filter by active_ids here; process everything and handle missing info later

            # 3. Create Harmonized Unified Data
            unified_rows = []
                
            # Add Transaction rows (Source: TRANS)
            if not df_trans.empty:
                for idx, row in df_trans.iterrows():
                    mat_id = self.normalize_id(row.get('MaterialID'))
                    mat_name = self.get_material_name_only(mat_id)
                    mat_info = self.get_material_info(mat_id)
                    
                    # [NEW] Filter: Only show consumables in In/Out history
                    if not self._is_consumable_material(mat_name, ''):
                        continue

                    # [NEW] Deduplication: Skip auto-generated transactions that are already covered by DAILY usage entries
                    # This removes the "50 vs -50" duplication while keeping the richer DAILY info.
                    note = str(row.get('Note', ''))
                    if "(자동 차감)" in note:
                        continue # Skip ALL auto-generated TRANS records as they are aggregated in the DAILY record

                    unified_rows.append({
                        'Date': row['Date'],
                        'Site': row.get('Site', ''),
                        'Type': row.get('Type', 'OUT'),
                        '품목명': mat_name,
                        'SN': mat_info.get('SN', ''),
                        '규격': mat_info.get('규격', ''),
                        'MaterialID': mat_id,
                        'Quantity': row.get('Quantity', 0),
                        'Warehouse': row.get('Warehouse', ''),
                        'User': row.get('User', ''),
                        'Note': row.get('Note', ''),
                        '차량번호': row.get('차량번호', ''),
                        'Source': 'TRANS',
                        'Fee': 0, # Transactions usually don't have inspection fees directly
                        'OrigIdx': idx
                    })


            # Add Daily Usage rows (Source: DAILY)
            if not df_daily.empty:
                for idx, row in df_daily.iterrows():
                    # [FLEXIBLE] Extract usage with multiple fallback names
                    u_val = _f(row.get('Usage', row.get('검사량', row.get('수량', row.get('Quantity', 0)))))
                    f_val = _f(row.get('FilmCount', row.get('매수', 0)))
                    
                    # If everything is 0, we still might want to show it if there's a note or site info
                    # but for now let's just make sure we capture NDT properly
                    ndt_vals_list = [
                        _f(row.get(f'NDT_{n}', row.get(n, 0))) 
                        for n in ['세척제', '침투제', '현상제', '백색페인트', '흑색자분', '형광자분', '형광침투제', '자분페인트']
                    ]
                    
                    # If any sign of life, keep the row
                    if u_val <= 0 and f_val <= 0 and not any(v > 0 for v in ndt_vals_list) and not str(row.get('Site', '')):
                        continue

                    mat_id = self.normalize_id(row.get('MaterialID'))
                    mat_name = self.get_material_name_only(mat_id)
                    mat_info = self.get_material_info(mat_id)
                    
                    # [NEW] Filter: Only show consumables in In/Out history (Exclude PAUT, etc.)
                    # If mat_name is empty (manual entry), check method or keywords in Note if needed, 
                    # but _is_consumable_material already handles method.
                    if not self._is_consumable_material(mat_name, row.get('검사방법', '')):
                        # Check if it has NDT chemicals (which ARE consumables)
                        if not any(v > 0 for v in ndt_vals_list):
                            continue
                    
                    # [FIX] Filter removed to show all site usage in history
                    
                    # Collect all worker names from User, User2...User10
                    workers = []
                    for i in range(1, 11):
                        u_col = 'User' if i == 1 else f'User{i}'
                        u_val_str = str(row.get(u_col, '')).strip()
                        if u_val_str and u_val_str not in ['nan', 'None', '']:
                            # Handle (주간) suffix if present
                            u_clean = re.sub(r'\(.*?\)\s*', '', u_val_str).strip()
                            if u_clean: workers.append(u_clean)
                    
                    worker_str = ", ".join(workers) if workers else str(row.get('User', ''))

                    # [FIX] Allow showing quantity for PT/MT/NDT items in history view
                    disp_qty = u_val

                    # [ROBUST] Try to parse ndt_data if it exists as a JSON string
                    ndt_json = {}
                    nj_raw = row.get('ndt_data', '')
                    if isinstance(nj_raw, str) and nj_raw.strip().startswith('{'):
                        try: ndt_json = json.loads(nj_raw)
                        except: pass
                    elif isinstance(nj_raw, dict):
                        ndt_json = nj_raw

                    unified_rows.append({
                        'Date': row['Date'],
                        'Site': row.get('Site', ''),
                        'Type': 'OUT',
                        '품목명': mat_name,
                        'SN': mat_info.get('SN', ''),
                        '규격': mat_info.get('규격', ''),
                        'MaterialID': mat_id,
                        'Quantity': disp_qty,
                        'NDT_세척제': _f(row.get('NDT_세척제', row.get('세척제', ndt_json.get('세척제', ndt_json.get('PT Cleaner', ndt_json.get('PTCleaner', 0)))))),
                        'NDT_침투제': _f(row.get('NDT_침투제', row.get('침투제', ndt_json.get('침투제', ndt_json.get('PT Penetrant', ndt_json.get('PTPenetrant', 0)))))),
                        'NDT_현상제': _f(row.get('NDT_현상제', row.get('현상제', ndt_json.get('현상제', ndt_json.get('PT Developer', ndt_json.get('PTDeveloper', 0)))))),
                        'NDT_백색페인트': _f(row.get('NDT_백색페인트', row.get('NDT_페인트', row.get('백색페인트', ndt_json.get('백색페인트', ndt_json.get('MT WHITE', ndt_json.get('MT7C-WHITE', 0))))))),
                        'NDT_흑색자분': row.get('NDT_흑색자분', row.get('NDT_자분', ndt_json.get('흑색자분', ndt_json.get('MT 7C-BLACK', 0)))),
                        'Warehouse': '',
                        'User': worker_str,
                        'Note': row.get('Note', ''),
                        '차량번호': row.get('차량번호', ''),
                        'Source': 'DAILY',
                        'Fee': row.get('검사비', 0.0), # Bring fee from daily usage
                        'OrigIdx': idx
                    })
            
            if not unified_rows:
                return
                
            df_unified = pd.DataFrame(unified_rows)
            
            # Apply Filters
            # Material Filter
            if hasattr(self, 'cb_trans_filter_mat'):
                selected_mat = str(self.cb_trans_filter_mat.get()).strip()
                if selected_mat and selected_mat != "전체":
                    # Check if MaterialID matches the display name in materials_df
                    matching_rows = []
                    for idx, row in df_unified.iterrows():
                        if self.get_material_display_name(row['MaterialID']) == selected_mat:
                            matching_rows.append(idx)
                    df_unified = df_unified.loc[matching_rows]
            
            # Site Filter
            if hasattr(self, 'cb_trans_filter_site'):
                selected_site = str(self.cb_trans_filter_site.get()).strip()
                if selected_site and selected_site != "전체":
                    df_unified = df_unified[df_unified['Site'].astype(str).str.contains(selected_site, na=False, case=False, regex=False)]
            
            # Vehicle Filter
            if hasattr(self, 'cb_trans_filter_vehicle'):
                selected_vehicle = str(self.cb_trans_filter_vehicle.get()).strip()
                if selected_vehicle and selected_vehicle != "전체":
                    df_unified = df_unified[df_unified['차량번호'].astype(str).str.contains(selected_vehicle, na=False, case=False, regex=False)]

            # [NEW] Update Filter Dropdowns with latest unified data unique values
            if hasattr(self, 'cb_trans_filter_site') and not getattr(self, '_site_filter_busy', False):
                self._site_filter_busy = True
                try:
                    s_current = self.cb_trans_filter_site.get()
                    s_vals = sorted(df_unified['Site'].dropna().astype(str).unique().tolist())
                    self.cb_trans_filter_site['values'] = ["전체"] + [v for v in s_vals if v.strip()]
                    if s_current: self.cb_trans_filter_site.set(s_current)
                    else: self.cb_trans_filter_site.set("전체")
                finally: self._site_filter_busy = False

            if hasattr(self, 'cb_trans_filter_mat') and not getattr(self, '_mat_filter_busy', False):
                self._mat_filter_busy = True
                try:
                    m_current = self.cb_trans_filter_mat.get()
                    # Get display names for all MaterialIDs in the unified set
                    unique_mat_ids = df_unified['MaterialID'].dropna().unique()
                    m_vals = sorted([self.get_material_display_name(mid) for mid in unique_mat_ids])
                    self.cb_trans_filter_mat['values'] = ["전체"] + [v for v in m_vals if v.strip()]
                    if m_current: self.cb_trans_filter_mat.set(m_current)
                    else: self.cb_trans_filter_mat.set("전체")
                finally: self._mat_filter_busy = False
                
            # [STABILITY] Ensure all standard columns are visible and reset any previous hiding
            all_inout_cols = ('날짜', '현장', '구분', '품목명', '수량', '세척제', '침투제', '현상제', '백색페인트', '흑색자분', 'SN', '규격', '창고', '담당자', '비고', '검사비')
            
            # Use saved configuration if available
            if hasattr(self, 'tab_config') and 'inout_visible_cols' in self.tab_config:
                self.inout_tree['displaycolumns'] = [c for c in self.tab_config['inout_visible_cols'] if c in all_inout_cols]
            else:
                self.inout_tree['displaycolumns'] = all_inout_cols

            # Sort and display
            df_sorted = df_unified.sort_values(by='Date', ascending=False).head(500)
            
            # [RESTORED] Column Order based on setup: 날짜, 구분, 품목명, SN, 규격, 수량, 현장, 창고, 담당자, 비고, 검사비
            for idx, row in df_sorted.iterrows():
                try:
                    # Robust date formatting with fallback for NaT
                    if pd.isna(row['Date']):
                        usage_date = "알 수 없음"
                    else:
                        usage_date = row['Date'].strftime('%Y-%m-%d %H:%M')
                    
                    # Manage Numeric Fee Formatting
                    f_val = row.get('Fee', 0.0)
                    fee_str = f"{f_val:,.0f}" if (isinstance(f_val, (int, float)) and f_val > 0) else ""

                    # Normalize Quantity for display: Use absolute value since 'Type' (IN/OUT) indicates direction
                    raw_qty = row.get('Quantity', row.get('수량', 0))
                    try:
                        clean_qty = abs(float(str(raw_qty).replace(',', '')))
                        qty_str = f"{clean_qty:.1f}" if clean_qty % 1 != 0 else f"{int(clean_qty)}"
                    except:
                        qty_str = self.clean_nan(raw_qty)

                    # NDT columns formatting
                    def _q_str(val):
                        try:
                            v = abs(float(str(val).replace(',', '')))
                            return f"{v:g}" if v > 0 else ""
                        except: return ""

                    self.inout_tree.insert('', tk.END, values=(
                        usage_date,
                        self.clean_nan(row.get('Site', '')),
                        self.clean_nan(row.get('Type', '-')),
                        self.clean_nan(row.get('품목명', '')),
                        qty_str,
                        _q_str(row.get('NDT_세척제', 0)),
                        _q_str(row.get('NDT_침투제', 0)),
                        _q_str(row.get('NDT_현상제', 0)),
                        _q_str(row.get('NDT_백색페인트', 0)),
                        _q_str(row.get('NDT_흑색자분', 0)),
                        self.clean_nan(row.get('SN', '')),
                        self.clean_nan(row.get('규격', '')),
                        self.clean_nan(row.get('Warehouse', '')),
                        self.clean_nan(row.get('User', '')),
                        self.clean_nan(row.get('Note', '')),
                        fee_str
                    ), tags=(str(row.get('OrigIdx', '')), row.get('Source', 'TRANS')))


                except Exception as e:
                    print(f"Row Display Error: {e}")
                    continue
                    
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Unified Transaction View Error: {e}\n{error_details}")

    def refresh_inout_history(self):
        """Refresh In/Out history and related filter source lists."""
        try:
            # Ensure filter candidates reflect latest data first
            self.update_material_combo()
            # Then redraw history and stock views
            self.update_transaction_view()
            self.update_stock_view()
        except Exception as e:
            messagebox.showerror("새로고침 오류", f"입출고 내역 새로고침 중 오류가 발생했습니다.\n{e}")

    def delete_transaction_entry(self):
        """Delete selected transaction from history and refresh"""
        selection = self.inout_tree.selection()
        if not selection:
            messagebox.showwarning("선택 오류", "삭제할 기록을 선택해주세요.")
            return
            
        if not messagebox.askyesno("삭제 확인", "선택한 기록을 삭제하시겠습니까?\n(삭제 시 재고 계산에 즉시 반영됩니다.)"):
            return
            
        for item in selection:
            tags = self.inout_tree.item(item, 'tags')
            if len(tags) >= 2:
                idx = int(tags[0])
                source = tags[1]
                
                if source == 'TRANS':
                    if idx in self.transactions_df.index:
                        self.transactions_df = self.transactions_df.drop(idx)
                elif source == 'DAILY':
                    if idx in self.daily_usage_df.index:
                        self.daily_usage_df = self.daily_usage_df.drop(idx)
                        # Optional: also delete from daily reports if needed
                        # But for now, removing from DailyUsage sheet is primary
        
        self.save_data()
        self.update_transaction_view()
        self.update_stock_view()
        
        # [FIX] Automatically refresh the Site tab and Query tab
        if hasattr(self, 'budget_view_tree'):
            self.update_budget_site_view()
        if hasattr(self, 'query_tree') and hasattr(self, 'cb_filter_year'):
            try:
                y = int(self.cb_filter_year.get().replace('년', ''))
                m = int(self.cb_filter_month.get().replace('월', ''))
                self.update_monthly_usage_view(y, m)
            except:
                pass
                
        messagebox.showinfo("완료", "거래 기록이 삭제되었습니다.")

    def on_material_selected(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import on_material_selected_impl
        return on_material_selected_impl(self, *args, **kwargs)

    def _on_trans_site_return(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import _on_trans_site_return_impl
        return _on_trans_site_return_impl(self, *args, **kwargs)

    def _on_warehouse_return(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import _on_warehouse_return_impl
        return _on_warehouse_return_impl(self, *args, **kwargs)

    def _on_user_return(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import _on_user_return_impl
        return _on_user_return_impl(self, *args, **kwargs)


    def format_worker_summary(self, *args, **kwargs):
        from site_apps.central.src.models.worker_model import format_worker_summary_impl
        return format_worker_summary_impl(self, *args, **kwargs)

    def show_worker_popup(self, event, tree):
        """Show a floating window (dropdown-style) with full record details for the clicked row"""
        item_id = tree.identify_row(event.y)
        column_id = tree.identify_column(event.x)
        if not item_id or not column_id: return
            
        col_name = tree.heading(column_id)['text']
        if col_name != '작업자': return
            
        tags = tree.item(item_id, 'tags')
        if not tags or 'total' in tags: return
            
        try:
            full_info = []
            if tree == getattr(self, 'daily_usage_tree', None):
                try:
                    df_idx = int(tags[0])
                    if df_idx not in self.daily_usage_df.index: return
                    entry = self.daily_usage_df.loc[df_idx]
                    
                    # 1. Basic Info
                    usage_date = self._safe_format_datetime(entry.get('Date', ''), '%Y-%m-%d')
                    full_info.append(f"[기본 기록 정보]")
                    full_info.append(f"• 날짜: {usage_date}")
                    full_info.append(f"• 현장: {self.clean_nan(entry.get('Site', ''))}")
                    
                    equip = self.clean_nan(entry.get('장비명', ''))
                    if equip: full_info.append(f"• 장비명: {equip}")
                    
                    method = self.clean_nan(entry.get('검사방법', ''))
                    if method: full_info.append(f"• 검사방법: {method}")
                    
                    amount = entry.get('검사량', 0)
                    if amount and float(amount) > 0: full_info.append(f"• 검사량: {amount}")
                    
                    
                    full_info.append(f"• 품목명: {self.get_material_display_name(entry.get('MaterialID'))}")
                    
                    # 2. Worker Info (All 10 slots)
                    full_info.append(f"\n[작업자 및 시간 정보]")
                    for i in range(1, 11):
                        u = self.clean_nan(entry.get('User' if i==1 else f'User{i}', ''))
                        if u:
                            t = self.clean_nan(entry.get('WorkTime' if i==1 else f'WorkTime{i}', ''))
                            o = self.clean_nan(entry.get('OT' if i==1 else f'OT{i}', ''))
                            full_info.append(f"• 작업자{i}: {u} | {t} | OT: {o}")
                    
                    # 3. Cost Info (Hide zeros)
                    cost_info = []
                    for k, label in [('단가', '단가'), ('출장비', '출장비'), ('일식', '일식'), ('검사비', '검사비')]:
                        raw_val = entry.get(k, 0)
                        try:
                            if pd.isna(raw_val) or str(raw_val).lower() == 'nan': continue
                            val_float = float(str(raw_val).replace(',', ''))
                            if val_float > 0:
                                cost_info.append(f"• {label}: {int(val_float):,}")
                        except: pass
                    if cost_info:
                        full_info.append(f"\n[비용 상세]")
                        full_info.extend(cost_info)
                    
                    # 4. RTK/NDT (Hide zeros)
                    rtk_found = []
                    rtk_cats = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타"]
                    for cat in rtk_cats:
                        val = entry.get(f'RTK_{cat}', 0)
                        if val and float(val) > 0:
                            rtk_found.append(f"  - {cat}: {val}")
                    if rtk_found:
                        full_info.append(f"\n[RT 결함 상세]")
                        full_info.extend(rtk_found)
                        
                    ndt_found = []
                    ndt_mats = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
                    for mat in ndt_mats:
                        val = entry.get(f'NDT_{mat}', 0)
                        if pd.isna(val): # Fallback
                            if mat == "흑색자분": val = entry.get('NDT_자분', 0)
                            elif mat == "백색페인트": val = entry.get('NDT_페인트', 0)
                        if val and float(val) > 0:
                            ndt_found.append(f"  - {mat}: {val}")
                    if ndt_found:
                        full_info.append(f"\n[NDT 자재 사용량]")
                        full_info.extend(ndt_found)
                    
                    note = entry.get('Note', '')
                    if note and str(note).strip(): full_info.append(f"\n• 비고: {note}")
                    
                    full_info.append(f"• 입력시간: {entry.get('EntryTime', '')}")
                    
                except Exception as e: 
                    print(f"Popup error: {e}")
                    return
            else:
                # Fallback for other trees (Monthly usage, etc.)
                vals = tree.item(item_id, 'values')
                cols = tree['columns']
                try:
                    w_idx = cols.index('작업자')
                    t_idx = cols.index('작업시간')
                    raw_names = vals[w_idx]
                    raw_times = vals[t_idx]
                    full_info.append(f"작업자: {raw_names}")
                    full_info.append(f"작업시간: {raw_times}")
                except: return

            if not full_info: return
            
            # Get cell bounding box for precise positioning
            bbox = tree.bbox(item_id, column_id)
            if not bbox: return
            
            x, y, w, h = bbox
            root_x = tree.winfo_rootx() + x
            root_y = tree.winfo_rooty() + y
            
            popup = tk.Toplevel(self.root)
            popup.overrideredirect(True)
            popup.attributes("-topmost", True)
            popup.configure(bg='white', highlightbackground="#0078d7", highlightthickness=2)
            
            popup_width = max(w, 550)
            popup_height = 450
            
            # 1. Horizontal positioning (Center relative to cell)
            final_x = root_x - (popup_width - w) // 2
            
            # 2. Vertical positioning (Show ABOVE the cell by default)
            final_y = root_y - popup_height - 5
            
            # [SMART CHECK] If the popup would go off-screen at the top, show it BELOW instead
            if final_y < 0:
                final_y = root_y + h + 5
            
            # Initial relative offsets from the cell root
            popup._rel_offset_x = final_x - root_x
            popup._rel_offset_y = final_y - root_y

            popup.geometry(f"{popup_width}x{popup_height}+{int(final_x)}+{int(final_y)}")
            
            # --- Custom Dragging Support ---
            def start_drag(event):
                popup._drag_data = {"x": event.x, "y": event.y}
            
            def do_drag(event):
                if hasattr(popup, '_drag_data'):
                    dx = event.x - popup._drag_data["x"]
                    dy = event.y - popup._drag_data["y"]
                    new_x = popup.winfo_x() + dx
                    new_y = popup.winfo_y() + dy
                    popup.geometry(f"+{new_x}+{new_y}")
                    
                    # Update relative offsets after manual drag so sync continues from new spot
                    try:
                        curr_bbox = tree.bbox(item_id, column_id)
                        if curr_bbox:
                            cx, cy, cw, ch = curr_bbox
                            # Important: Update relative to the CURRENT root position
                            popup._rel_offset_x = new_x - (tree.winfo_rootx() + cx)
                            popup._rel_offset_y = new_y - (tree.winfo_rooty() + cy)
                    except: pass
            
            popup.bind("<Button-1>", start_drag)
            popup.bind("<B1-Motion>", do_drag)
            
            content_frame = ttk.Frame(popup, padding=10)
            content_frame.pack(fill='both', expand=True)
            
            lb = tk.Listbox(content_frame, font=('Malgun Gothic', 10), 
                            bg='white', fg='#333333', relief='flat', 
                            borderwidth=0, highlightthickness=0)
            for info in full_info:
                lb.insert(tk.END, info)
            
            lb.pack(side='left', fill='both', expand=True)
            
            sb = ttk.Scrollbar(content_frame, orient="vertical", command=lb.yview)
            lb.configure(yscrollcommand=sb.set)
            sb.pack(side='right', fill='y')
            
            popup.focus_set()
            
            # --- Window Move Synchronization ---
            def reposition_popup(event=None):
                if not popup.winfo_exists(): return
                # Re-calculate position based on CURRENT root position + saved relative offset
                try:
                    curr_bbox = tree.bbox(item_id, column_id)
                    if not curr_bbox: return
                    cx, cy, cw, ch = curr_bbox
                    croot_x = tree.winfo_rootx() + cx
                    croot_y = tree.winfo_rooty() + cy
                    
                    # Apply relative offset to keep the sync accurate
                    new_x = croot_x + popup._rel_offset_x
                    new_y = croot_y + popup._rel_offset_y
                    
                    popup.geometry(f"+{int(new_x)}+{int(new_y)}")
                except: pass
                
            # Bind to root move/resize
            bind_id = self.root.bind("<Configure>", reposition_popup, add="+")
            
            def close_popup(e=None):
                if popup.winfo_exists():
                    try: self.root.unbind("<Configure>", bind_id)
                    except: pass
                    popup.destroy()
            
            popup.bind("<FocusOut>", close_popup)
            popup.bind("<Escape>", close_popup)
            popup.bind("<Double-1>", close_popup)
            
            # [FIX] Removed global self.root.bind("<Button-1>") which leaked and interfered with DateEntry.
            # FocusOut is sufficient for closing the popup when clicking elsewhere.
        except Exception as e:
            print(f"Final popup error: {e}")
    def clean_df_export(self, df):
        """Sanitize data for Excel: remove illegal characters, handle empty values, and protect formulas"""
        # 1. Pre-clean common empty representations
        df = df.replace(['nan', 'NaN', 'None'], "")
        
        # 2. Identify columns that have at least one meaningful value
        def is_really_empty(col):
            # A column is empty if all stripped values are "", NaN, or 0 sequences
            return df[col].astype(str).str.strip().replace(['nan', 'None', '', '0', '0.0', '0.00'], pd.NA).dropna().empty

        non_empty_cols = [col for col in df.columns if not is_really_empty(col)]
        
        # Always keep essential columns even if empty
        essential = ['날짜', '현장', '품목명', 'Date', 'Site']
        # Also keep standard RTK and NDT result columns
        rtk_cats = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타", "RTK총계"]
        ndt_materials = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
        essential.extend(rtk_cats)
        essential.extend([f'RTK_{c}' for c in rtk_cats])
        essential.extend(ndt_materials)
        essential.extend([f'NDT_{"".join(m.split())}' for m in ndt_materials])
        
        for col in essential:
            if col in df.columns and col not in non_empty_cols:
                non_empty_cols.append(col)
                
        # 3. Sanitize content and protect formulas
        def sanitize_and_escape(val):
            if pd.isna(val) or val == "" or val == 0 or val == 0.0:
                return ""
            
            # String conversion and cleaning
            s_val = str(val).strip()
            if not s_val or s_val.lower() in ['nan', 'none']:
                return ""

            # Remove illegal XML control characters (ASCII < 32 except tab, LF, CR)
            # These characters cause Excel to report file corruption.
            s_val = "".join(ch for ch in s_val if ord(ch) >= 32 or ch in "\t\n\r")
            
            # Escape potential formula characters (=, +, -, @)
            # IMPORTANT: Skip escaping if the value is a valid number (e.g., -5000)
            if s_val.startswith(('=', '+', '-', '@')):
                try:
                    # If it's a number, don't prefix with '
                    float(s_val)
                    return val
                except ValueError:
                    # It's a text string starting with a formula char
                    return f"'{s_val}"
            
            return s_val

        # Apply sanitization to all columns
        for col in df.columns:
            df[col] = df[col].apply(sanitize_and_escape)
        
        # Return only non-empty columns in original order
        final_cols = [c for c in df.columns if c in non_empty_cols]
        return df[final_cols]

    def save_df_to_excel_autofit(self, *args, **kwargs):
        if "export" in "save_df_to_excel_autofit" or "excel" in "save_df_to_excel_autofit":
            from site_apps.central.src.services.excel_exporter import save_df_to_excel_autofit_impl
        else:
            from site_apps.central.src.services.data_loader import save_df_to_excel_autofit_impl
        return save_df_to_excel_autofit_impl(self, *args, **kwargs)


    
    def setup_import_tab(self):
        import_frame = ttk.LabelFrame(self.tab_import, text="데이터 관리")
        import_frame.pack(pady=20, padx=20, fill='both', expand=True)
        
        # Import Section
        ttk.Label(import_frame, text="엑셀 파일에서 자재 데이터 가져오기", font=('Arial', 11, 'bold')).pack(pady=10)
        ttk.Label(import_frame, text="형식: MaterialID, 회사코드, 관리품번, 품목명, 창고, 모델명, 규격, 품목군코드, 제조사, 제조국, 가격, 관리단위, 수량", 
                 wraplength=600).pack(pady=5)
        
        btn_import = ttk.Button(import_frame, text="엑셀 파일 가져오기", command=self.import_from_excel)
        btn_import.pack(pady=10)
        
        ttk.Separator(import_frame, orient='horizontal').pack(fill='x', pady=20)
        
        # Export Section
        ttk.Label(import_frame, text="현재 데이터 엑셀로 내보내기", font=('Arial', 11, 'bold')).pack(pady=10)
        
        btn_export_materials = ttk.Button(import_frame, text="자재 목록 내보내기", command=self.export_materials)
        btn_export_materials.pack(pady=5)
        
        btn_export_trans = ttk.Button(import_frame, text="거래 내역 내보내기", command=self.export_transactions)
        btn_export_trans.pack(pady=5)
        
        btn_export_all = ttk.Button(import_frame, text="전체 데이터 내보내기", command=self.export_all)
        btn_export_all.pack(pady=5)
    
    def import_from_excel(self):
        file_path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if not file_path:
            return
        
        try:
            # Load raw data
            raw_imported_df = pd.read_excel(file_path)
            
            # 1. Standardize column names (mapping) summerly
            col_map = {
                '품목명': ['품목명', '품명', 'Item Name', 'Name'],
                'SN': ['SN', 'SN번호', 'Serial Number', 'S/N'],
                '규격': ['규격', 'Specification', 'Spec'],
                '수량': ['수량', 'Initial Stock', 'Current Stock', 'Qty', 'Amount'],
                '모델명': ['모델명', 'Model', 'Machine'],
                '창고': ['창고', 'Warehouse', 'Location'],
                '가격': ['가격', 'Price', 'Unit Price'],
                '제조사': ['제조사', 'Manufacturer', 'Brand'],
                '공급업체': ['공급업체', '공급업자', 'Supplier', 'Vendor'],
                '관리단위': ['관리단위', 'Unit', 'UOM'],
                '재고하한': ['재고하한', '재주문 수준', 'Reorder Point', 'Min Stock']
            }
            
            # Simple column normalization
            working_df = pd.DataFrame()
            for std_name, aliases in col_map.items():
                for alias in aliases:
                    if alias in raw_imported_df.columns:
                        working_df[std_name] = raw_imported_df[alias]
                        break
                if std_name not in working_df.columns:
                    if std_name in ['품목명', '수량']: # Critical columns defaults
                        if std_name == '수량': working_df[std_name] = 0
                        else: 
                            messagebox.showerror("오류", f"필수 컬럼 '{std_name}'을(를) 찾을 수 없습니다.")
                            return
                    else:
                        working_df[std_name] = ""

            # Standardize values: clean strings, handle NaNs summerly
            working_df['품목명'] = working_df['품목명'].astype(str).str.strip()
            working_df['SN'] = working_df['SN'].astype(str).replace('nan', '').str.strip()
            working_df['규격'] = working_df['규격'].astype(str).replace('nan', '').str.strip()
            working_df['수량'] = pd.to_numeric(working_df['수량'], errors='coerce').fillna(0)
            working_df['가격'] = pd.to_numeric(working_df['가격'], errors='coerce').fillna(0)
            
            # [Added] Extract SN from Model name if necessary summerly
            def _extract_row_sn(row):
                m, s = self.extract_sn_from_model(row['모델명'], row['SN'])
                return pd.Series([m, s])
            
            working_df[['모델명', 'SN']] = working_df.apply(_extract_row_sn, axis=1)
            
            # Filter out rows without a name
            working_df = working_df[working_df['품목명'] != 'nan']
            working_df = working_df[working_df['품목명'] != '']
            
            # 2. Interior Deduplication: merge same items within the Excel file itself summerly
            # We group by Name, SN, and Spec.
            final_excel_df = working_df.groupby(['품목명', 'SN', '규격'], as_index=False).agg({
                '수량': 'sum',
                '가격': 'max',
                '모델명': 'first',
                '창고': 'first',
                '제조사': 'first',
                '공급업체': 'first',
                '관리단위': 'first',
                '재고하한': 'first'
            })
            
            # 3. Preparation: Match with existing inventory
            count_new = 0
            count_merged = 0
            count_overwritten = 0
            count_skipped = 0
            
            duplicate_indices = [] # (excel_row_idx, db_row_idx)
            new_items = []
            
            for idx, row in final_excel_df.iterrows():
                mat_name, sn, spec = row['품목명'], row['SN'], row['규격']
                
                existing_idx = -1
                if not self.materials_df.empty:
                    mask = (self.materials_df['품목명'].astype(str) == mat_name) & \
                           (self.materials_df['SN'].astype(str).replace('nan', '') == sn) & \
                           (self.materials_df['규격'].astype(str).replace('nan', '') == spec)
                    matches = self.materials_df.index[mask].tolist()
                    if matches:
                        existing_idx = matches[0]
                
                if existing_idx != -1:
                    duplicate_indices.append((idx, existing_idx))
                else:
                    new_items.append(idx)
            
            # 4. Conflict Resolution Choice Dialog summerly
            mode = "SKIP"
            if duplicate_indices:
                choices_dlg = tk.Toplevel(self.root)
                choices_dlg.title("중복 항목 처리 선택")
                choices_dlg.geometry("450x300")
                choices_dlg.transient(self.root)
                choices_dlg.grab_set()
                
                selection = tk.StringVar(value="CANCEL")
                
                ttk.Label(choices_dlg, text=f"기존 재고 목록과 일치하는 항목 {len(duplicate_indices)}건이 발견되었습니다.", 
                          font=('Malgun Gothic', 10, 'bold')).pack(pady=15)
                ttk.Label(choices_dlg, text="어떻게 처리할까요?").pack(pady=5)
                
                def set_choice(c):
                    selection.set(c)
                    choices_dlg.destroy()
                
                btn_frame = ttk.Frame(choices_dlg, padding=10)
                btn_frame.pack(fill='both', expand=True)
                
                ttk.Button(btn_frame, text="기존 재고에 수량 합치기 (추천)", 
                           command=lambda: set_choice("MERGE"), width=40).pack(pady=5)
                ttk.Button(btn_frame, text="기존 정보 덮어쓰기 (Excel 정보로 교체)", 
                           command=lambda: set_choice("OVERWRITE"), width=40).pack(pady=5)
                ttk.Button(btn_frame, text="중복 항목 건너뛰기 (신규만 추가)", 
                           command=lambda: set_choice("SKIP"), width=40).pack(pady=5)
                ttk.Button(btn_frame, text="가져오기 취소", 
                           command=lambda: set_choice("CANCEL"), width=40).pack(pady=15)
                
                self.root.wait_window(choices_dlg)
                mode = selection.get()
                if mode == "CANCEL":
                    return

            # 5. Process Import based on Mode summerly
            # Add new items first
            new_rows = []
            for idx in new_items:
                row_data = final_excel_df.iloc[idx].to_dict()
                new_mat_id = self.materials_df['MaterialID'].max() + 1 if not self.materials_df.empty else 1
                row_data['MaterialID'] = new_mat_id
                row_data['Active'] = 1
                new_rows.append(row_data)
                count_new += 1
            
            if new_rows:
                self.materials_df = pd.concat([self.materials_df, pd.DataFrame(new_rows)], ignore_index=True)
            
            # Handle duplicates
            if mode != "SKIP":
                for ex_idx, db_idx in duplicate_indices:
                    ex_row = final_excel_df.iloc[ex_idx]
                    
                    if mode == "MERGE":
                        # Add to existing quantity
                        current_qty = pd.to_numeric(self.materials_df.at[db_idx, '수량'], errors='coerce')
                        if pd.isna(current_qty): current_qty = 0
                        self.materials_df.at[db_idx, '수량'] = current_qty + ex_row['수량']
                        # Also update Manufacturer/Supplier if empty in DB but present in Excel
                        for field in ['제조사', '공급업체', '규격', '창고']:
                            if field in self.materials_df.columns and field in ex_row:
                                if not self.materials_df.at[db_idx, field] or str(self.materials_df.at[db_idx, field]) == 'nan':
                                    self.materials_df.at[db_idx, field] = ex_row[field]
                        count_merged += 1
                        
                    elif mode == "OVERWRITE":
                        # Full replace (except ID and Active status)
                        for col in final_excel_df.columns:
                            if col in self.materials_df.columns:
                                self.materials_df.at[db_idx, col] = ex_row[col]
                        count_overwritten += 1
            else:
                count_skipped = len(duplicate_indices)

            # 6. Finalize summerly
            self.save_data()
            self.update_material_combo()
            self.update_stock_view()
            self.update_registration_combos()
            
            summary = f"자재 가져오기가 완료되었습니다.\n\n"
            summary += f"• 신규 등록: {count_new}건\n"
            if mode == "MERGE":
                summary += f"• 재고량 합산: {count_merged}건\n"
            elif mode == "OVERWRITE":
                summary += f"• 정보 덮어쓰기: {count_overwritten}건\n"
            elif mode == "SKIP":
                summary += f"• 중복 건너뜀: {count_skipped}건\n"
            
            messagebox.showinfo("가져오기 완료", summary)
            
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messagebox.showerror("오류", f"파일을 가져오는데 실패했습니다: {e}")
            
        except Exception as e:
            messagebox.showerror("오류", f"파일을 가져오는데 실패했습니다: {e}")
    
    def export_materials(self, *args, **kwargs):
        if "export" in "export_materials" or "excel" in "export_materials":
            from site_apps.central.src.services.excel_exporter import export_materials_impl
        else:
            from site_apps.central.src.services.data_loader import export_materials_impl
        return export_materials_impl(self, *args, **kwargs)
    
    def export_transactions(self):
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="Transactions_Export.xlsx",
            title="거래 내역 저장",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if save_path:
            try:
                self.transactions_df.to_excel(save_path, index=False)
                messagebox.showinfo("완료", "거래 내역이 저장되었습니다.")
            except Exception as e:
                messagebox.showerror("오류", f"저장 실패: {e}")
    
    def export_transaction_history(self):
        """Export the detailed In/Out history to Excel summerly"""
        history_data = []
        columns = [self.inout_tree.heading(col, 'text') for col in self.inout_tree['columns']]
        
        for item in self.inout_tree.get_children():
            values = self.inout_tree.item(item, 'values')
            row_data = {}
            for i, col in enumerate(columns):
                row_data[col] = values[i] if i < len(values) else ''
            history_data.append(row_data)
        
        if not history_data:
            messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
            return

    def _migrate_worktimes(self, input_list):
        """Helper to standardize work times with 익일 marker and sort logically"""
        default_worktimes = [
            "09:00~18:00", "09:00~19:00", "09:00~20:00", "09:00~21:00", 
            "09:00~22:00", "09:00~23:00", "09:00~24:00", "09:00~익일01:00",
            "09:00~익일02:00", "09:00~익일03:00", "18:00~익일02:00", "18:00~익일03:00"
        ]
        
        processed_times = set()
        for t in list(input_list) + default_worktimes:
            if not t or '~' not in t: continue
            
            clean_t = str(t).replace('익일', '').strip()
            try:
                start, end = clean_t.split('~')
                s_h = int(start.split(':')[0])
                e_h = int(end.split(':')[0])
                
                if e_h < s_h and '익일' not in str(t):
                    processed_times.add(f"{start}~익일{end}")
                else:
                    processed_times.add(str(t).strip())
            except:
                processed_times.add(str(t).strip())

        def time_sort_key(t):
            has_ikil = 1 if '익일' in t else 0
            clean = t.replace('익일', '')
            return (has_ikil, clean)

        return sorted(list(processed_times), key=time_sort_key)
        
        # Prepare filename
        today = datetime.datetime.now().strftime('%Y%m%d')
        filename = f"입출고내역_{today}.xlsx"
        
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=filename,
            title="입출고 내역 저장",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if save_path:
            try:
                history_df = pd.DataFrame(history_data)
                history_df = self.clean_df_export(history_df)
                self.save_df_to_excel_autofit(history_df, save_path, "입출고내역")
                messagebox.showinfo("완료", "입출고 내역이 저장되었습니다.")
            except Exception as e:
                messagebox.showerror("오류", f"저장 실패: {e}")
    
    def export_all(self):
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="Complete_Export.xlsx",
            title="전체 데이터 저장",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if save_path:
            try:
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    self.materials_df.to_excel(writer, sheet_name='Materials', index=False)
                    self.transactions_df.to_excel(writer, sheet_name='Transactions', index=False)
                messagebox.showinfo("완료", "전체 데이터가 저장되었습니다.")
            except Exception as e:
                messagebox.showerror("오류", f"저장 실패: {e}")
    

    def view_monthly_usage(self):
        """Display monthly usage by item name in the treeview"""
        # Clear current view
        for item in self.usage_tree.get_children():
            self.usage_tree.delete(item)
        
        year = int(self.cb_year.get())
        month = int(self.cb_month.get())
        
        # Filter transactions for the selected month
        month_mask = (self.transactions_df['Date'].dt.year == year) & \
                     (self.transactions_df['Date'].dt.month == month) & \
                     (self.transactions_df['Type'] == 'OUT')
        monthly_trans = self.transactions_df[month_mask]
        
        # Filter transactions for cumulative (from start of year to selected month)
        cumulative_mask = (self.transactions_df['Date'].dt.year == year) & \
                          (self.transactions_df['Date'].dt.month <= month) & \
                          (self.transactions_df['Type'] == 'OUT')
        cumulative_trans = self.transactions_df[cumulative_mask]
        
        # Build usage data for each material
        usage_data = []
        for _, mat in self.materials_df.iterrows():
            mat_id = mat['MaterialID']
            
            # Calculate monthly usage
            month_usage = monthly_trans[monthly_trans['MaterialID'] == mat_id]['Quantity'].sum()
            
            # Calculate cumulative usage (year-to-date)
            cumulative_usage = cumulative_trans[cumulative_trans['MaterialID'] == mat_id]['Quantity'].sum()
            
            # Only show items with usage
            if month_usage > 0 or cumulative_usage > 0:
                usage_data.append({
                    '품목명': mat.get('품목명', ''),  # Using 재고현황's 품목명 field
                    '관리품번': mat.get('관리품번', ''),
                    '규격': mat.get('규격', ''),
                    '단위': mat.get('관리단위', 'EA'),
                    '월사용량': month_usage,
                    '누계사용량': cumulative_usage
                })
        
        # Sort by item name
        usage_data.sort(key=lambda x: x['품목명'])
        
        # Display in treeview
        for data in usage_data:
            self.usage_tree.insert('', tk.END, values=(
                data['품목명'],
                data['관리품번'],
                data['규격'],
                data['단위'],
                f"{data['월사용량']:.1f}",
                f"{data['누계사용량']:.1f}"
            ))
        
        # Show message if no data
        if not usage_data:
            messagebox.showinfo("알림", f"{year}년 {month}월에 사용 내역이 없습니다.")

    def generate_yearly_report(self):
        year = int(self.cb_year.get())
        # Filter transactions for the year
        mask = (self.transactions_df['Date'].dt.year == year) & (self.transactions_df['Type'] == 'OUT')
        yearly_trans = self.transactions_df[mask]
        
        report = []
        for _, mat in self.materials_df.iterrows():
            mat_id = mat['MaterialID']
            
            # Calculate monthly usage for each month
            row_data = {
                '설비코드': mat.get('관리품번', ''),
                '자재명': mat.get('품목명', ''),  # Using 재고현황's 품목명 field
                '분류': mat.get('품목군코드', ''),
                '규격': mat.get('규격', ''),
                '단위': mat.get('관리단위', ''),
                '제조사': mat.get('제조사', '')
            }
            
            # Add monthly columns (1월 ~ 12월)
            monthly_values = []
            for month in range(1, 13):
                month_mask = (yearly_trans['MaterialID'] == mat_id) & \
                            (yearly_trans['Date'].dt.month == month)
                month_usage = yearly_trans[month_mask]['Quantity'].sum()
                row_data[f'{month}월'] = month_usage
                monthly_values.append(month_usage)
            
            # Calculate totals
            total = sum(monthly_values)
            row_data['합계'] = total
            
            # Calculate cumulative total
            cumulative = 0
            for i, val in enumerate(monthly_values, 1):
                cumulative += val
                if i == 12:  # Only show final cumulative at the end
                    row_data['누계'] = cumulative
            
            report.append(row_data)
            
        report_df = pd.DataFrame(report)
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                 initialfile=f"Yearly_Usage_{year}.xlsx",
                                                 title="보고서 저장")
        if save_path:
            report_df.to_excel(save_path, index=False)
            messagebox.showinfo("완료", f"{year}년 보고서가 저장되었습니다.")

    def generate_monthly_report(self):
        year = int(self.cb_year.get())
        month = int(self.cb_month.get())
        
        mask = (self.transactions_df['Date'].dt.year == year) & \
               (self.transactions_df['Date'].dt.month == month) & \
               (self.transactions_df['Type'] == 'OUT')
        monthly_trans = self.transactions_df[mask]
        
        report = []
        for _, mat in self.materials_df.iterrows():
            mat_id = mat['MaterialID']
            total_usage = monthly_trans[monthly_trans['MaterialID'] == mat_id]['Quantity'].sum()
            report.append({
                '설비코드': mat.get('관리품번', ''),
                '자재명': mat.get('품목명', ''),  # Using 재고현황's 품목명 field
                '분류': mat.get('품목군코드', ''),
                '규격': mat.get('규격', ''),
                '단위': mat.get('관리단위', ''),
                '제조사': mat.get('제조사', ''),
                '월간 총 사용량': total_usage
            })
            
        report_df = pd.DataFrame(report)
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                 initialfile=f"Monthly_Usage_{year}_{month}.xlsx",
                                                 title="보고서 저장")
        if save_path:
            report_df.to_excel(save_path, index=False)
            messagebox.showinfo("완료", f"{year}년 {month}월 보고서가 저장되었습니다.")

    def setup_monthly_usage_tab(self, *args, **kwargs):
        from site_apps.central.src.views.monthly_usage_view import setup_monthly_usage_tab_impl
        return setup_monthly_usage_tab_impl(self, *args, **kwargs)
    
    def update_monthly_usage_view(self, *args, **kwargs):
        from site_apps.central.src.views.monthly_usage_view import update_monthly_usage_view_impl
        return update_monthly_usage_view_impl(self, *args, **kwargs)

    def on_monthly_usage_select(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import on_monthly_usage_select_impl
        return on_monthly_usage_select_impl(self, *args, **kwargs)

    def _populate_monthly_summary_trees(self, df, has_note=None):
        """Helper to fill site and worker summary trees with given data subset"""
        # Clear current views
        for item in self.site_summary_tree.get_children():
            self.site_summary_tree.delete(item)
        for item in self.worker_summary_tree.get_children():
            self.worker_summary_tree.delete(item)
            
        # [NEW] Clear detached views
        detached = self.detached_windows.get('monthly')
        if detached:
            for item in detached['site_tree'].get_children(): detached['site_tree'].delete(item)
            for item in detached['worker_tree'].get_children(): detached['worker_tree'].delete(item)
            
        if df.empty:
            return

        # --- Populate Site Summary ---
        # Include MaterialID in grouping and add relevant NDT/Cost fields to aggregation
        site_agg_dict = {
            '검사비': 'sum',
            '출장비': 'sum',
            '제경비': 'sum',
            '기술료': 'sum',
            '환산물량': 'sum',
            '재료비': 'sum',
            '인건비': 'sum',
            'Usage': 'sum',
            '검사량': 'sum'
        }
        
        # Add RTK fields for RT매수 calculation
        rtk_fields = ['RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크', 'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타']
        for rf in rtk_fields:
            if rf in df.columns: 
                # [DEFENSIVE] Force numeric before aggregation
                df[rf] = pd.to_numeric(df[rf], errors='coerce').fillna(0.0)
                site_agg_dict[rf] = 'sum'
            
        # Add NDT fields
        ndt_map = {
            'NDT_형광자분': 'sum',
            'NDT_자분': 'sum',
            'NDT_흑색자분': 'sum',
            'NDT_페인트': 'sum',
            'NDT_백색페인트': 'sum',
            'NDT_침투제': 'sum',
            'NDT_세척제': 'sum',
            'NDT_현상제': 'sum',
            'NDT_형광': 'sum',
            'NDT_형광침투제': 'sum'
        }
        for nf in ndt_map:
            if nf in df.columns: 
                # [DEFENSIVE] Force numeric before aggregation
                df[nf] = pd.to_numeric(df[nf], errors='coerce').fillna(0.0)
                site_agg_dict[nf] = 'sum'

        # Helper for joining values in site-only summary
        def join_unique_non_empty(series):
            vals = [" ".join(str(v).split()) for v in series if pd.notna(v) and str(v).strip()]
            return " | ".join(sorted(set(vals)))
            
        # Add Joiner for mixed MaterialIDs in site-only summary
        site_agg_dict['MaterialID'] = join_unique_non_empty
        
        # Ensure Site and Method are strings and filled
        if 'Site' in df.columns: df['Site'] = df['Site'].fillna('').astype(str)
        if '검사방법' in df.columns: df['검사방법'] = df['검사방법'].fillna('미지정').astype(str)
        
        site_summary = df.groupby(['Site', '검사방법'], dropna=False).agg(site_agg_dict).reset_index()
        
        # [NEW] Track active columns for dynamic hiding
        site_cols = ('현장', '검사방법', '품목명', '수량', '검사비', '출장비', '제경비', '기술료', '환산물량', '재료비', '인건비', '형광자분', '흑색자분', '백색페인트', 
                     '침투제', '세척제', '현상제', '형광침투제', '센터미스', '농도', '마킹미스', '필름마크', '취급부주의', '고객불만', '기타', 'RTK총계')
        active_cols = set(['현장', '검사방법', '품목명', '수량']) # Mandatory columns
        
        # Define a robust active check for numeric data
        def is_active(val):
            if val is None: return False
            s = str(val).strip().lower()
            if s in ('', '0', '0.0', '0.00', 'nan', 'none', '-'):
                return False
            try:
                # Remove common separators and units for numeric check
                v = float(s.replace(',', '').replace('원', '').replace('시간', '').strip())
                return abs(v) > 0.001
            except:
                return bool(s)

        for _, row in site_summary.iterrows():
            # Get RT film usage from unified Usage column
            film_total = row.get('Usage', 0.0)
            
            # Calculate values
            black_mag = row.get('NDT_자분', 0.0) + row.get('NDT_흑색자분', 0.0)
            white_paint = row.get('NDT_페인트', 0.0) + row.get('NDT_백색페인트', 0.0)
            pt_pen = row.get('NDT_침투제', 0.0)
            pt_cln = row.get('NDT_세척제', 0.0)
            pt_dev = row.get('NDT_현상제', 0.0)
            pt_fluoro = row.get('NDT_형광', 0.0) + row.get('NDT_형광침투제', 0.0)
            
            r_center = row.get('RTK_센터미스', 0.0)
            r_density = row.get('RTK_농도', 0.0)
            r_marking = row.get('RTK_마킹미스', 0.0)
            r_film = row.get('RTK_필름마크', 0.0)
            r_careless = row.get('RTK_취급부주의', 0.0)
            r_customer = row.get('RTK_고객불만', 0.0)
            r_other = row.get('RTK_기타', 0.0)
            rtk_sum = r_center + r_density + r_marking + r_film + r_careless + r_customer + r_other
            
            # Track which columns are active using robust threshold
            if is_active(row.get('검사비', 0.0)): active_cols.add('검사비')
            if is_active(row.get('출장비', 0.0)): active_cols.add('출장비')
            if is_active(row.get('제경비', 0.0)): active_cols.add('제경비')
            if is_active(row.get('기술료', 0.0)): active_cols.add('기술료')
            if is_active(row.get('환산물량', 0.0)): active_cols.add('환산물량')
            if is_active(row.get('재료비', 0.0)): active_cols.add('재료비')
            if is_active(row.get('인건비', 0.0)): active_cols.add('인건비')
            if is_active(row.get('NDT_형광자분', 0.0)): active_cols.add('형광자분')
            if is_active(black_mag): active_cols.add('흑색자분')
            if is_active(white_paint): active_cols.add('백색페인트')
            if is_active(pt_pen): active_cols.add('침투제')
            if is_active(pt_cln): active_cols.add('세척제')
            if is_active(pt_dev): active_cols.add('현상제')
            if is_active(pt_fluoro): active_cols.add('형광침투제')
            if is_active(r_center): active_cols.add('센터미스')
            if is_active(r_density): active_cols.add('농도')
            if is_active(r_marking): active_cols.add('마킹미스')
            if is_active(r_film): active_cols.add('필름마크')
            if is_active(r_careless): active_cols.add('취급부주의')
            if is_active(r_customer): active_cols.add('고객불만')
            if is_active(r_other): active_cols.add('기타')
            if is_active(rtk_sum): active_cols.add('RTK총계')
            
            # Get Material Name (Handle multiple joined IDs if site grouping consolidated them)
            mat_ids_raw = str(row['MaterialID']).split(' | ')
            mat_names_list = []
            for m_id_str in mat_ids_raw:
                curr_id = m_id_str.strip()
                if not curr_id or curr_id.lower() == 'nan': continue
                
                # [ROBUST] Logic to find material name from ID (handling numeric conversion)
                def get_single_mat_name(m_id_val):
                    try: 
                        m_f = float(m_id_val)
                        matches = self.materials_df[pd.to_numeric(self.materials_df['MaterialID'], errors='coerce') == m_f]
                        if not matches.empty: return str(matches.iloc[0]['품목명'])
                    except: pass
                    matches = self.materials_df[self.materials_df['MaterialID'].astype(str) == str(m_id_val)]
                    if not matches.empty: return str(matches.iloc[0]['품목명'])
                    return f"ID:{m_id_val}"
                
                mat_names_list.append(get_single_mat_name(curr_id))
                    
            mat_name = ", ".join(sorted(set(mat_names_list))) if mat_names_list else "품목 미지정"

            values = (
                row['Site'],
                row['검사방법'],
                mat_name,
                f"{row.get('검사량', 0.0):.1f}" if is_active(row.get('검사량', 0.0)) else '',
                f"{row.get('검사비', 0.0):,.0f}" if is_active(row.get('검사비', 0.0)) else '',
                f"{row.get('출장비', 0.0):,.0f}" if is_active(row.get('출장비', 0.0)) else '',
                f"{row.get('제경비', 0.0):,.0f}" if is_active(row.get('제경비', 0.0)) else '',
                f"{row.get('기술료', 0.0):,.0f}" if is_active(row.get('기술료', 0.0)) else '',
                f"{row.get('환산물량', 0.0):.1f}" if is_active(row.get('환산물량', 0.0)) else '',
                f"{row.get('재료비', 0.0):,.0f}" if is_active(row.get('재료비', 0.0)) else '',
                f"{row.get('인건비', 0.0):,.0f}" if is_active(row.get('인건비', 0.0)) else '',
                f"{row.get('NDT_형광자분', 0.0):.1f}" if is_active(row.get('NDT_형광자분', 0.0)) else '',
                f"{black_mag:.1f}" if is_active(black_mag) else '',
                f"{white_paint:.1f}" if is_active(white_paint) else '',
                f"{pt_pen:.1f}" if is_active(pt_pen) else '',
                f"{pt_cln:.1f}" if is_active(pt_cln) else '',
                f"{pt_dev:.1f}" if is_active(pt_dev) else '',
                f"{pt_fluoro:.1f}" if is_active(pt_fluoro) else '',
                f"{r_center:.1f}" if is_active(r_center) else '',
                f"{r_density:.1f}" if is_active(r_density) else '',
                f"{r_marking:.1f}" if is_active(r_marking) else '',
                f"{r_film:.1f}" if is_active(r_film) else '',
                f"{r_careless:.1f}" if is_active(r_careless) else '',
                f"{r_customer:.1f}" if is_active(r_customer) else '',
                f"{r_other:.1f}" if is_active(r_other) else '',
                f"{rtk_sum:.1f}" if is_active(rtk_sum) else ''
            )
            self.site_summary_tree.insert('', tk.END, values=values)
            if detached:
                detached['site_tree'].insert('', tk.END, values=values)

        # Apply dynamic column hiding to Site Summary
        visible_sum_cols = [col for col in site_cols if col in active_cols]
        self.site_summary_tree['displaycolumns'] = visible_sum_cols
        if detached:
            detached['site_tree']['displaycolumns'] = visible_sum_cols

        # --- Populate Worker Summary ---
        worker_data = []
        
        # [ROBUST] Use the same precise column detection as update_monthly_usage_view
        def find_paired_cols_local(cols):
            pairs = []
            col_set = set(cols)
            for i in range(1, 11):
                u_n = 'User' if i == 1 else f'User{i}'
                w_n = 'WorkTime' if i == 1 else f'WorkTime{i}'
                o_n = 'OT' if i == 1 else f'OT{i}'
                if u_n in col_set:
                    pairs.append((u_n, w_n if w_n in col_set else None, o_n if o_n in col_set else None))
            return pairs

        df_norm = df.copy()
        df_norm.columns = [str(c).strip().replace(' ', '') for c in df_norm.columns]
        
        pairs = find_paired_cols_local(df_norm.columns)
        for uc, wc, oc in pairs:
            temp_df = pd.DataFrame()
            temp_df['WorkerName'] = df_norm[uc]
            temp_df['ShiftType'] = df_norm[wc] if wc else '주간'
            temp_df['OTValue'] = df_norm[oc] if oc else ''
            
            # Context for deduping
            temp_df['Date'] = df_norm['Date'] if 'Date' in df_norm.columns else None
            temp_df['Site'] = df_norm['Site'] if 'Site' in df_norm.columns else ''
            temp_df['MaterialID'] = df_norm['MaterialID'] if 'MaterialID' in df_norm.columns else ''
            temp_df['WorkTime'] = df_norm['WorkTime'] if 'WorkTime' in df_norm.columns else '' # Base worktime for deduping
            
            worker_data.append(temp_df)
        
        if worker_data:
            # [NEW] Track active columns for dynamic hiding in Worker Summary
            worker_sum_cols = ('작업자', '총공수', '연장(시간)', '야간(시간)', '휴일(시간)', '총OT(시간)', '연장(금액)', '야간(금액)', '휴일(금액)', '총OT(금액)')
            active_worker_cols = set(['작업자', '총공수'])
            
            worker_df = pd.concat(worker_data)
            worker_df['WorkerName'] = worker_df['WorkerName'].apply(self.clean_nan)
            worker_df = worker_df[worker_df['WorkerName'] != '']
            
            # [NEW] Worker-level Deduping to prevent double-counting in split records
            if not worker_df.empty:
                def make_worker_key(row):
                    d = self._safe_format_datetime(row.get('Date', ''), '%Y-%m-%d')
                    s = str(row.get('Site', '')).strip()
                    w = str(row.get('WorkTime', '')).strip()
                    m = str(row.get('MaterialID', '')).strip()
                    un = str(row.get('WorkerName', '')).strip()
                    return (d, s, w, m, un)
                
                worker_df['_dedupe_key'] = worker_df.apply(make_worker_key, axis=1)
                # Keep the first instance of a worker in any given activity unit
                worker_df = worker_df.drop_duplicates(subset=['_dedupe_key']).drop(columns=['_dedupe_key'])
            
            if not worker_df.empty:
                # [FIXED] Use row['Date'] if available, otherwise fallback to first available date in subset
                def get_date_val(row):
                    if 'Date' in row and pd.notna(row['Date']): return row['Date']
                    try: return df.iloc[0]['Date']
                    except: return None

                # [NEW] Enhanced OT hour extraction: check WorkTime if OTValue is only amount
                def get_ot_hours_robust(row):
                    ot_val = str(row['OTValue']).strip()
                    if not ot_val or ot_val == '0': return 0.0
                    
                    # If it's already "N시간...", parse it
                    if '시간' in ot_val:
                        return self._parse_ot_hours(ot_val)
                    
                    # If it's just an amount, calculate hours from WorkTime
                    if ot_val.replace(',', '').isdigit() and int(ot_val.replace(',', '')) > 100:
                        wt_val = row['ShiftType']
                        date_val = get_date_val(row)
                        h, _ = self._calculate_ot_from_worktime(wt_val, date_val)
                        return h
                    
                    return self._parse_ot_hours(ot_val)

                worker_df['OT_H'] = worker_df.apply(get_ot_hours_robust, axis=1)
                
                # Split calculation also needs to be robust
                def get_split_robust(row):
                    ot_val = str(row['OTValue']).strip()
                    date_val = get_date_val(row)
                    if '시간' in ot_val or not (ot_val.replace(',', '').isdigit() and int(ot_val.replace(',', '')) > 100):
                        return self._calculate_split_ot_hours(ot_val, date_val)
                    else:
                        # Recalculate from WorkTime
                        wt_val = row['ShiftType']
                        
                        # We need the simulation logic from _calculate_split_ot_hours but using wt_val
                        # Let's use a temporary string that _calculate_split_ot_hours can parse
                        h, _ = self._calculate_ot_from_worktime(wt_val, date_val)
                        fake_ot_str = f"{h}시간 ({ot_val}원)"
                        # Check start hour from wt_val for range logic
                        marker_pattern = MARKER_PATTERN
                        clean_wt = marker_pattern.sub('', str(wt_val)).strip()
                        if '~' in clean_wt:
                            start_time = clean_wt.split('~')[0]
                            fake_ot_str = f"{start_time}~{h}시간 ({ot_val}원)"
                        
                        return self._calculate_split_ot_hours(fake_ot_str, date_val)

                # Now returns 3 values (day, night, holiday_dawn)
                worker_df[['OT_H_Day', 'OT_H_Night', 'OT_H_Holiday_Dawn']] = worker_df.apply(
                    lambda r: pd.Series(get_split_robust(r)), axis=1
                )
                worker_df['OT_A'] = worker_df['OTValue'].apply(self.calculate_ot_amount)
                
                # [FIXED] Parse fractional work counts from ShiftType (e.g., "0.5주간" -> 0.5)
                def parse_work_count(val):
                    s = str(val).strip()
                    if not s: return 1.0
                    try:
                        # Extract first float/int found in string
                        match = _re.search(r'([0-9\.]+)', s)
                        if match: return float(match.group(1))
                    except: pass
                    return 1.0
                
                worker_df['Count'] = worker_df['ShiftType'].apply(parse_work_count)
                
                # [FIXED] Shift types: Automatically treat Saturday/Sunday as Holiday shifts
                def determine_shift_holiday(row):
                    s_type = str(row['ShiftType']).strip()
                    if '휴일' in s_type: return 1.0
                    d_val = get_date_val(row)
                    if d_val:
                        try:
                            if pd.to_datetime(d_val).weekday() >= 5: return 1.0
                        except: pass
                    return 0.0

                worker_df['Shift_Holiday'] = worker_df.apply(determine_shift_holiday, axis=1)
                worker_df['Shift_Day'] = worker_df['ShiftType'].apply(lambda x: 1.0 if '주간' in str(x) and '휴일' not in str(x) else 0.0)
                worker_df['Shift_Night'] = worker_df['ShiftType'].apply(lambda x: 1.0 if '야간' in str(x) and '휴일' not in str(x) else 0.0)
                
                # If explicit holiday shift, all OT is holiday.
                # If NOT explicit holiday shift, add the OT_H_Holiday_Dawn (Friday dawn) to holiday hours.
                worker_df['H_Holiday'] = worker_df.apply(lambda r: r['OT_H'] if r['Shift_Holiday'] > 0 else r['OT_H_Holiday_Dawn'], axis=1)
                
                # Calculate amount properly. If it was fully holiday, it's just OT_A.
                # If it was Friday Dawn, add the calculated dawn hours * 7500 to whatever other OT they had?
                # Actually, A_Holiday should just be the amount from those hours
                worker_df['A_Holiday'] = worker_df.apply(lambda r: r['OT_A'] if r['Shift_Holiday'] > 0 else r['OT_H_Holiday_Dawn'] * 7500, axis=1)
                
                # For day/night, they only get values if it's NOT a full holiday shift.
                worker_df['H_Day'] = worker_df.apply(lambda r: r['OT_H_Day'] if r['Shift_Holiday'] == 0 else 0.0, axis=1)
                worker_df['H_Night'] = worker_df.apply(lambda r: r['OT_H_Night'] if r['Shift_Holiday'] == 0 else 0.0, axis=1)
                worker_df['A_Day'] = worker_df.apply(lambda r: r['H_Day'] * 4000 if r['Shift_Holiday'] == 0 else 0.0, axis=1)
                worker_df['A_Night'] = worker_df.apply(lambda r: r['H_Night'] * 5000 if r['Shift_Holiday'] == 0 else 0.0, axis=1)

                # [REVERTED] Back to entry-based count
                worker_summary = worker_df.groupby('WorkerName').agg({
                    'Count': 'sum',
                    'H_Day': 'sum',
                    'H_Night': 'sum',
                    'H_Holiday': 'sum',
                    'OT_H': 'sum',
                    'A_Day': 'sum',
                    'A_Night': 'sum',
                    'A_Holiday': 'sum',
                    'OT_A': 'sum'
                }).reset_index()
                
                for _, row in worker_summary.iterrows():
                    # Track active columns using robust is_active helper
                    if is_active(row['H_Day']): active_worker_cols.add('연장(시간)')
                    if is_active(row['H_Night']): active_worker_cols.add('야간(시간)')
                    if is_active(row['H_Holiday']): active_worker_cols.add('휴일(시간)')
                    if is_active(row['OT_H']): active_worker_cols.add('총OT(시간)')
                    if is_active(row['A_Day']): active_worker_cols.add('연장(금액)')
                    if is_active(row['A_Night']): active_worker_cols.add('야간(금액)')
                    if is_active(row['A_Holiday']): active_worker_cols.add('휴일(금액)')
                    if is_active(row['OT_A']): active_worker_cols.add('총OT(금액)')
 
                    values = (
                        row['WorkerName'],
                        f"{row['Count']:.1f}" if is_active(row['Count']) else "",
                        f"{row['H_Day']:.1f}" if is_active(row['H_Day']) else "",
                        f"{row['H_Night']:.1f}" if is_active(row['H_Night']) else "",
                        f"{row['H_Holiday']:.1f}" if is_active(row['H_Holiday']) else "",
                        f"{row['OT_H']:.1f}" if is_active(row['OT_H']) else "",
                        f"{row['A_Day']:,.0f}" if is_active(row['A_Day']) else "",
                        f"{row['A_Night']:,.0f}" if is_active(row['A_Night']) else "",
                        f"{row['A_Holiday']:,.0f}" if is_active(row['A_Holiday']) else "",
                        f"{row['OT_A']:,.0f}" if is_active(row['OT_A']) else ""
                    )
                    self.worker_summary_tree.insert('', tk.END, values=values)
                    if detached:
                        detached['worker_tree'].insert('', tk.END, values=values)
                
                # Apply dynamic column hiding to Worker Summary
                visible_worker_cols = [col for col in worker_sum_cols if col in active_worker_cols]
                self.worker_summary_tree['displaycolumns'] = visible_worker_cols
                if detached:
                    detached['worker_tree']['displaycolumns'] = visible_worker_cols

    def apply_worker_shift_hours_to_budget(self):
        """월별 탭의 작업자별 누계(주간/야간/휴일) 시간을 공사실행예산서 특별근무 투입시간에 적용한다."""
        if not hasattr(self, 'labor_detail_widget'):
            messagebox.showwarning("공사탭 미초기화",
                                   "공사실행예산서 탭을 먼저 열어 초기화하세요.")
            return

        # worker_summary_tree에서 모든 행의 주간/야간/휴일 합산
        total_h_day = 0.0
        total_h_night = 0.0
        total_h_holiday = 0.0
        worker_count_day = 0
        worker_count_night = 0
        worker_count_holiday = 0

        def _f(val):
            try: return float(str(val).replace(',', '') or 0)
            except: return 0.0

        for item in self.worker_summary_tree.get_children():
            vals = self.worker_summary_tree.item(item, 'values')
            # 컬럼 순서: 작업자, 총공수, 연장(시간), 야간(시간), 휴일(시간), 총OT(시간), 연장(금액), 야간(금액), 휴일(금액), 총OT(금액)
            h_day     = _f(vals[2]) if len(vals) > 2 else 0.0
            h_night   = _f(vals[3]) if len(vals) > 3 else 0.0
            h_holiday = _f(vals[4]) if len(vals) > 4 else 0.0
            if h_day > 0:
                total_h_day += h_day
                worker_count_day += 1
            if h_night > 0:
                total_h_night += h_night
                worker_count_night += 1
            if h_holiday > 0:
                total_h_holiday += h_holiday
                worker_count_holiday += 1

        if total_h_day == 0 and total_h_night == 0 and total_h_holiday == 0:
            messagebox.showinfo("데이터 없음",
                                "작업자별 누계에 주간/야간/휴일 시간 데이터가 없습니다.\n"
                                "월별 탭에서 조회를 먼저 실행하세요.")
            return

        # LaborCostDetailWidget 특별근무 섹션에 입력
        ldw = self.labor_detail_widget
        # 연장근무 ← 주간 OT 시간 합계 / 인원수
        if worker_count_day > 0:
            avg_day = total_h_day / worker_count_day
            ldw.entries["연장근무"]['personnel'].delete(0, 'end')
            ldw.entries["연장근무"]['personnel'].insert(0, f"{worker_count_day:g}")
            ldw.entries["연장근무"]['period'].delete(0, 'end')
            ldw.entries["연장근무"]['period'].insert(0, f"{avg_day:.1f}")
        # 야간근무 ← 야간 OT 시간 합계
        if worker_count_night > 0:
            avg_night = total_h_night / worker_count_night
            ldw.entries["야간근무"]['personnel'].delete(0, 'end')
            ldw.entries["야간근무"]['personnel'].insert(0, f"{worker_count_night:g}")
            ldw.entries["야간근무"]['period'].delete(0, 'end')
            ldw.entries["야간근무"]['period'].insert(0, f"{avg_night:.1f}")
        # 휴일근무 ← 휴일 OT 시간 합계
        if worker_count_holiday > 0:
            avg_holiday = total_h_holiday / worker_count_holiday
            ldw.entries["휴일근무"]['personnel'].delete(0, 'end')
            ldw.entries["휴일근무"]['personnel'].insert(0, f"{worker_count_holiday:g}")
            ldw.entries["휴일근무"]['period'].delete(0, 'end')
            ldw.entries["휴일근무"]['period'].insert(0, f"{avg_holiday:.1f}")

        # 계산 반영
        ldw.calculate_all()

        # 공사실행예산서 탭으로 이동
        try:
            tab_idx = [self.notebook.tab(i, 'text') for i in range(self.notebook.index('end'))].index('공사실행예산서')
            self.notebook.select(tab_idx)
        except:
            pass

        messagebox.showinfo("적용 완료",
                            f"특별근무 투입시간 적용 완료!\n"
                            f"  연장근무: {worker_count_day}명 / {total_h_day:.1f}h (평균 {total_h_day/max(worker_count_day,1):.1f}h/인)\n"
                            f"  야간근무: {worker_count_night}명 / {total_h_night:.1f}h (평균 {total_h_night/max(worker_count_night,1):.1f}h/인)\n"
                            f"  휴일근무: {worker_count_holiday}명 / {total_h_holiday:.1f}h (평균 {total_h_holiday/max(worker_count_holiday,1):.1f}h/인)")

    def export_monthly_usage_history(self):
        """Export monthly usage data, site summaries, and worker summaries to a multi-sheet Excel file"""
        try:
            # 1. Collect Main Monthly Usage Data
            columns = self.monthly_usage_tree['columns']
            monthly_data = []
            for item in self.monthly_usage_tree.get_children():
                monthly_data.append(self.monthly_usage_tree.item(item, 'values'))
            
            if not monthly_data:
                messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
                return

            # 2. Collect Site Summary Data
            site_columns = self.site_summary_tree['columns']
            site_data = []
            for item in self.site_summary_tree.get_children():
                site_data.append(self.site_summary_tree.item(item, 'values'))

            # 3. Collect Worker Summary Data
            worker_columns = self.worker_summary_tree['columns']
            worker_data = []
            for item in self.worker_summary_tree.get_children():
                worker_data.append(self.worker_summary_tree.item(item, 'values'))
            
            # Prepare filename
            today = datetime.datetime.now().strftime('%Y%m%d')
            # Using a descriptive filename indicating it's a comprehensive report (종합)
            filename = f"월별집계_종합_{today}.xlsx"
            
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=filename,
                title="월별 집계(종합) 내역 저장",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not save_path:
                return

            # --- Process Monthly Usage Tree (Sheet 1) ---
            columns_to_export = [c for c in columns if c != '(Full작업자)']
            col_to_idx = {col: i for i, col in enumerate(columns)}
            
            final_monthly_data = []
            for row in monthly_data:
                new_row = list(row)
                # Swap summarized worker with full list if (Full작업자) exists
                if '작업자' in col_to_idx and '(Full작업자)' in col_to_idx:
                    full_idx = col_to_idx['(Full작업자)']
                    if full_idx < len(new_row):
                        full_val = new_row[full_idx]
                        if full_val:
                            new_row[col_to_idx['작업자']] = full_val
                
                # Filter only display columns
                final_row = [new_row[i] for i, c in enumerate(columns) if c != '(Full작업자)' and i < len(new_row)]
                final_monthly_data.append(final_row)
            
            monthly_df = pd.DataFrame(final_monthly_data, columns=columns_to_export)
            monthly_df = self.clean_df_export(monthly_df)

            # --- Process Site Summary (Sheet 2) ---
            site_df = pd.DataFrame(site_data, columns=site_columns)
            site_df = self.clean_df_export(site_df)

            # --- Process Worker Summary (Sheet 3) ---
            worker_df = pd.DataFrame(worker_data, columns=worker_columns)
            worker_df = self.clean_df_export(worker_df)

            # --- Save to Excel with Multiple Sheets and AutoFit ---
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                # Local helper for autofitting to avoid polluting global namespace
                def write_sheet(df, sheet_name):
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    worksheet = writer.sheets[sheet_name]
                    
                    def get_display_width(s):
                        width = 0
                        for char in str(s):
                            if ord(char) > 127: width += 2 # Double width for Korean
                            else: width += 1
                        return width

                    for idx, col in enumerate(df.columns):
                        series = df[col].astype(str)
                        lengths = series.apply(get_display_width)
                        max_val_len = lengths.max() if not lengths.empty else 0
                        header_len = get_display_width(col)
                        # Cap at 40 to allow better horizontal density in summaries
                        final_width = min(max(max_val_len, header_len) + 2, 40)
                        
                        col_letter = worksheet.cell(row=1, column=idx+1).column_letter
                        worksheet.column_dimensions[col_letter].width = final_width

                write_sheet(monthly_df, "월별집계_내역")
                write_sheet(site_df, "현장별_누계")
                write_sheet(worker_df, "작업자별_누계")

            messagebox.showinfo("완료", "월별 집계 내역(종합)이 저장되었습니다.\n(내역, 현장별, 작업자별 3개 시트 포함)")
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패: {e}")
            traceback.print_exc()

    def create_draggable_container(self, parent, label_text, widget_class, config_key, manage_list_key=None, grid_info=None, **widget_kwargs):
        """Create a draggable container with a label and a widget, styled as a visible box"""
        box_border_color = (
            getattr(self, 'style', None).lookup('TLabelframe', 'bordercolor') if hasattr(self, 'style') else ''
        ) or (
            getattr(self, 'style', None).lookup('TFrame', 'bordercolor') if hasattr(self, 'style') else ''
        ) or '#c0c0c0'

        container = tk.Frame(
            parent,
            relief="solid",
            borderwidth=1,
            highlightthickness=1,
            highlightbackground=box_border_color,
            bg=self.theme_bg
        )
        
        # Header container for label and buttons
        hdr = ttk.Frame(container) 
        hdr.pack(side='top', fill='x', padx=0, pady=0)
        
        # Drag handle icon
        lbl_drag = ttk.Label(hdr, text="✥", font=('Arial', 9), cursor='fleur')
        lbl_drag.pack(side='left', padx=1)
        self.make_header_draggable(lbl_drag, container)
        
        # Label
        lbl = ttk.Label(hdr, text=label_text, font=('Malgun Gothic', 9, 'bold'))
        lbl.pack(side='left', padx=1)
        self.make_header_draggable(lbl, container)
        
        # Icons container on the right
        btn_box = ttk.Frame(hdr)
        btn_box.pack(side='right', padx=1)

        # Rename icon
        btn_rename = ttk.Label(btn_box, text="✏️", font=('Arial', 8), cursor='hand2')
        btn_rename.pack(side='left', padx=1)
        btn_rename.bind('<Button-1>', lambda e: self.rename_widget_label(config_key))
        
        # Clone icon
        btn_clone = ttk.Label(btn_box, text="📋", font=('Arial', 8), cursor='hand2')
        btn_clone.pack(side='left', padx=1)
        btn_clone.bind('<Button-1>', lambda e: self.clone_widget(config_key))

        # Delete icon (X)
        btn_del = ttk.Label(btn_box, text="❌", font=('Arial', 8), cursor='hand2')
        btn_del.pack(side='left', padx=1)
        btn_del.bind('<Button-1>', lambda e: self.remove_box(config_key))

        # Internal Content Area
        content_area = ttk.Frame(container, padding=(1, 0))
        content_area.pack(side='top', fill='both', expand=True)

        # Widget
        widget = widget_class(content_area, **widget_kwargs)
        widget.pack(side='left', fill='both', expand=True)

        # Manage List Icon (Gear) - if it's a list-based widget
        if manage_list_key:
            btn_manage = ttk.Label(btn_box, text="⚙️", font=('Arial', 8), cursor='hand2')
            btn_manage.pack(side='left', padx=1)
            # [FIX] Bind AFTER widget is defined
            btn_manage.bind('<Button-1>', lambda e, cb=widget: self.open_list_management_dialog(manage_list_key, target_cb=cb))
        
        # If the widget is a basic tk widget (Text, Canvas), set its background
        if hasattr(widget, 'config') and 'bg' in widget.keys():
            try: widget.config(bg=self.theme_bg)
            except: pass
        
        # NEW: Handle grid placement first if grid_info provided
        if grid_info:
            container.grid(**grid_info)
        
        # Track for layout reset/config
        container._config_key = config_key
        container._label_widget = lbl
        container._widget = widget
        container._widget_class = widget_class
        container._widget_kwargs = widget_kwargs
        container._manage_list_key = manage_list_key
        
        # Register and make draggable (this now captures the CORRECT grid info)
        self.draggable_items[config_key] = container
        self.make_draggable(container, config_key)
        
        return container, widget

    def open_list_management_dialog(self, title_or_key, data_list=None, config_key=None, target_cb=None):
        """Open a generic dialog to manage (edit/delete) items in a data list"""
        if self.layout_locked: return

        # If data_list is None, we assume title_or_key is the config_key
        if data_list is None:
            config_key = title_or_key
            data_map = {
                'sites': ('현장 목록 관리', self.sites),
                'users': ('담당자 목록 관리', getattr(self, 'users', [])),
                'equipments': ('장비 목록 관리', getattr(self, 'equipments', [])),
                'vehicles': ('차량 목록 관리', getattr(self, 'vehicles', [])),
                'companies': ('업체 목록 관리', getattr(self, 'companies', [])),
                'materials': ('품목 목록 관리 (기본)', getattr(self, 'carestream_films', [])),
                'daily_units': ('단위 목록 관리', self.daily_units),
                'test_items': ('검사품명 목록 관리', getattr(self, 'test_items', [])),
                'applied_codes': ('적용코드 목록 관리', getattr(self, 'applied_codes', []))
            }
            if config_key not in data_map: return
            title, data_list = data_map[config_key]
        else:
            # Traditional 3-argument call: title, data_list, config_key
            title = title_or_key
            if config_key is None: config_key = title # Fallback

        # ── 이미 열린 창이 있으면 앞으로 가져오고 종료 ──────────────
        if not hasattr(self, '_list_mgmt_dialogs'):
            self._list_mgmt_dialogs = {}
        existing = self._list_mgmt_dialogs.get(config_key)
        if existing and existing.winfo_exists():
            existing.lift()
            existing.focus_set()
            return

        dialog = tk.Toplevel(self.root)
        self._list_mgmt_dialogs[config_key] = dialog  # 창 추적 등록
        dialog.title(title)
        dialog.geometry("600x450")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding=10)
        frame.pack(fill='both', expand=True)
        
        # [NEW] Search box for the list
        search_f = ttk.Frame(frame)
        search_f.pack(fill='x', pady=(0, 5))
        ttk.Label(search_f, text="🔍:").pack(side='left', padx=2)
        search_var = tk.StringVar()
        search_ent = ttk.Entry(search_f, textvariable=search_var)
        search_ent.pack(side='left', fill='x', expand=True, padx=2)
        search_ent.focus_set()

        # Add scrollbar
        scrollbar = ttk.Scrollbar(frame)
        scrollbar.pack(side='right', fill='y')
        
        listbox = tk.Listbox(frame, font=('Arial', 10))
        listbox.pack(fill='both', expand=True, side='left')
        listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=listbox.yview)
        
        def refresh_list(*args):
            query = search_var.get().strip().lower()
            listbox.delete(0, 'end')
            # Use original data_list for filtering
            for item in sorted(data_list):
                if not query or query in item.lower():
                    listbox.insert('end', item)
            # Trigger app-wide update of related comboboxes
            self.refresh_ui_for_list_change(config_key)

        search_var.trace_add("write", refresh_list)
        refresh_list()
            
        def select_and_close():
            sel = listbox.curselection()
            if not sel: return
            val = listbox.get(sel[0])
            if target_cb:
                # [FIX] Support both Entry and Combobox widgets
                if hasattr(target_cb, 'set'):
                    target_cb.set(val)
                else:
                    target_cb.delete(0, tk.END)
                    target_cb.insert(0, val)
                # Manually trigger any search/filter logic if needed
                if hasattr(self, 'update_daily_usage_view'):
                    self.update_daily_usage_view()
            dialog.destroy()

        listbox.bind("<Double-1>", lambda e: select_and_close())
            
        btn_frame = ttk.Frame(dialog, padding=5)
        btn_frame.pack(fill='x')
        
        # [NEW] Select button (Shows only if target_cb is present, or always as an alternative to double-click)
        ttk.Button(btn_frame, text="✅ 선택", command=select_and_close).pack(side='left', padx=2)

        def add_item():
            new_val = simpledialog.askstring("추가", "새 이름을 입력하세요:")
            if new_val and new_val.strip():
                val = new_val.strip()
                if val not in data_list:
                    data_list.append(val)
                    data_list.sort()
                    self.save_tab_config()
                    refresh_list()
                else:
                    messagebox.showinfo("정보", "이미 목록에 있는 이름입니다.")

        def edit_item():
            sel = listbox.curselection()
            if not sel: return
            idx = sel[0]
            old_val = listbox.get(idx)
            new_val = simpledialog.askstring("수정", f"[{old_val}] -> 새 이름을 입력하세요:", initialvalue=old_val)
            if new_val and new_val.strip() and new_val != old_val:
                if old_val in data_list: data_list.remove(old_val)
                data_list.append(new_val.strip())
                data_list.sort()
                self.save_tab_config()
                refresh_list()

        def delete_item():
            sel = listbox.curselection()
            if not sel: return
            idx = sel[0]
            val = listbox.get(idx)
            if messagebox.askyesno("삭제 확인", f"'{val}'을 목록에서 삭제하시겠습니까?"):
                if val in data_list: data_list.remove(val)
                # [NEW] If the deleted value is currently selected in the target combobox, clear it
                if target_cb:
                    current_val = target_cb.get().strip()
                    if current_val == val.strip():
                        if hasattr(target_cb, 'set'):
                            target_cb.set('')
                        else:
                            target_cb.delete(0, tk.END)
                self.save_tab_config()
                refresh_list()

        def on_close():
            self._list_mgmt_dialogs.pop(config_key, None)
            dialog.destroy()

        dialog.protocol("WM_DELETE_WINDOW", on_close)

        def bulk_register_to_stock():
            if messagebox.askyesno("일괄 등록 확인", "현재 목록에 있는 모든 품목을 '재고 현황' 데이터베이스에 새로 등록하시겠습니까?\n(이미 등록된 품목은 중복 등록되지 않습니다.)"):
                self.bulk_register_materials_to_stock(data_list)

        ttk.Button(btn_frame, text="추가", command=add_item).pack(side='left', padx=5, expand=True)
        ttk.Button(btn_frame, text="수정", command=edit_item).pack(side='left', padx=5, expand=True)
        ttk.Button(btn_frame, text="삭제", command=delete_item).pack(side='left', padx=5, expand=True)
        
        if config_key == 'materials':
            ttk.Button(btn_frame, text="📦 재고 일괄 등록", command=bulk_register_to_stock).pack(side='left', padx=5, expand=True)
            
        ttk.Button(btn_frame, text="닫기", command=on_close).pack(side='left', padx=5, expand=True)

    def bulk_register_materials_to_stock(self, material_list):
        """Register items from the preferred list to the main Material inventory if they don't exist or are inactive"""
        print(f"DEBUG: Starting bulk registration for {len(material_list)} items...")
        if not material_list:
            messagebox.showwarning("경고", "등록할 품목이 목록에 없습니다.")
            return
            
        added_count = 0
        reactivated_count = 0
        existing_indices = {} # Name.lower() -> index in materials_df
        max_id = 10000
        
        # Ensure materials_df is loaded
        if self.materials_df is None:
            self.materials_df = pd.DataFrame(columns=['MaterialID', '품목명', 'Active'])

        if not self.materials_df.empty:
            # Collect existing names and their positions
            # [STABILITY] Build index using (Name, Model) to allow parent items even if sub-items exist
            existing_keys = {} # (name_lower, model_lower) -> index
            for idx, mat in self.materials_df.iterrows():
                nm = str(mat.get('품목명', '')).strip().lower()
                md = str(mat.get('모델명', '')).strip().lower()
                if nm:
                    existing_keys[(nm, md)] = idx
            
            try:
                # Calculate next MaterialID
                if 'MaterialID' in self.materials_df.columns:
                    valid_ids = pd.to_numeric(self.materials_df['MaterialID'], errors='coerce').dropna()
                    if not valid_ids.empty:
                        max_id = int(valid_ids.max())
            except Exception as e:
                print(f"DEBUG: Error calculating max_id: {e}")
                max_id = 10000

        new_rows = []
        for mat in material_list:
            mat_clean = str(mat).strip()
            if not mat_clean: continue
            
            mat_lower = mat_clean.lower()
            # Check for parent item (empty model)
            if (mat_lower, '') in existing_keys:
                idx = existing_keys[(mat_lower, '')]
                # Check if it's inactive
                curr_active = self.materials_df.at[idx, 'Active']
                if str(curr_active) == '0' or curr_active == 0:
                    # Reactivate!
                    self.materials_df.at[idx, 'Active'] = 1
                    reactivated_count += 1
                    print(f"DEBUG: Reactivated item: {mat_clean}")
                else:
                    print(f"DEBUG: Skipping '{mat_clean}' (Already Active)")
            else:
                # Add as New
                max_id += 1
                new_row = {
                    'MaterialID': max_id,
                    '품목명': mat_clean,
                    'SN': '',
                    '모델명': '',
                    '관리단위': '매',
                    '수량': 0,
                    '재고하한': 0,
                    'Active': 1,
                    '구분': '소모품'
                }
                for col in self.materials_df.columns:
                    if col not in new_row:
                        new_row[col] = ''
                
                new_rows.append(new_row)
                added_count += 1
                print(f"DEBUG: Prepared new item: {mat_clean} (ID: {max_id})")
        
        if new_rows or reactivated_count > 0:
            try:
                if new_rows:
                    new_df = pd.DataFrame(new_rows)
                    new_df = new_df.reindex(columns=self.materials_df.columns, fill_value='')
                    self.materials_df = pd.concat([self.materials_df, new_df], ignore_index=True)
                
                # Save and Refresh
                self.save_data()
                self.update_stock_view()
                self.update_material_combo()
                
                msg = ""
                if added_count > 0: msg += f"{added_count}개의 품목 신규 등록 완료\n"
                if reactivated_count > 0: msg += f"{reactivated_count}개의 품목 재활성화(복구) 완료"
                
                messagebox.showinfo("등록 완료", msg)
            except Exception as e:
                print(f"DEBUG: Error during bulk registration: {e}")
                messagebox.showerror("오류", f"일괄 등록 중 오류가 발생했습니다:\n{e}")
        else:
            messagebox.showinfo("정보", "이미 모든 품목이 재고 현황에 활성화되어 있습니다.")

    def refresh_ui_for_list_change(self, config_key):
        """Update all related UI elements after a list (sites, users, etc) has changed"""
        # Dictionary mapping config keys to their current values
        list_map = {
            'sites': self.sites,
            'users': self.users,
            'warehouses': self.warehouses,
            'equipments': self.equipments,
            'worktimes': self.worktimes,
            'vehicles': self.vehicles,
            'companies': getattr(self, 'companies', []),
            'daily_units': getattr(self, 'daily_units', [])
        }
        
        if config_key not in list_map:
            return
            
        current_vals = list_map[config_key]
        # [NEW] Ensure uniqueness and strip whitespace before sorting
        unique_vals = sorted(list(set([str(v).strip() for v in current_vals if v])))
        sorted_vals = unique_vals
        
        # [FIX] Update the list in-place to ensure all UI dialogs holding references see the changes
        if isinstance(current_vals, list):
            current_vals[:] = sorted_vals
        
        # 1. Update standard widgets
        if config_key == 'sites':
            if hasattr(self, 'cb_daily_site'): self.cb_daily_site['values'] = sorted_vals
            if hasattr(self, 'cb_trans_site'): self.cb_trans_site['values'] = sorted_vals
            if hasattr(self, 'cb_daily_filter_site'):
                self.cb_daily_filter_site['values'] = ['전체'] + sorted_vals
            if hasattr(self, 'cb_budget_site'): self.cb_budget_site['values'] = sorted_vals
            if hasattr(self, 'cb_budget_view_site'): self.cb_budget_view_site['values'] = sorted_vals
        elif config_key == 'companies':
            # [NEW] Update all company comboboxes in daily usage
            if hasattr(self, 'cb_daily_company'): self.cb_daily_company['values'] = sorted_vals
            if hasattr(self, 'cb_daily_filter_company'):
                self.cb_daily_filter_company['values'] = ['전체'] + sorted_vals
        elif config_key == 'daily_units':
            print(f"DEBUG: Refreshing All Unit Comboboxes with: {sorted_vals}")
            # [FIX] daily_units is already updated in-place above via current_vals[:] = sorted_vals
            # 1. Re-merge to include DB units and update all UI automatically
            self.update_registration_combos()
            # update_registration_combos() handles cb_daily_unit and cb_unit automatically
        elif config_key == 'users':
            # Updated to match current attribute names if needed, 
            # but usually it's cb_daily_user for 1, and cb_daily_user{i} for 2-10
            for i in range(1, 11):
                attr = 'cb_daily_user' if i == 1 else f'cb_daily_user{i}'
                if hasattr(self, attr):
                    widget = getattr(self, attr)
                    if isinstance(widget, WorkerCompositeWidget):
                         widget.cb_name['values'] = [''] + sorted_vals
                    elif hasattr(widget, 'configure'):
                        try: widget['values'] = [''] + sorted_vals
                        except: pass
            
            if hasattr(self, 'ent_user'): self.ent_user['values'] = [''] + sorted_vals
        elif config_key == 'worktimes':
            # Update all worktime fields (1-10)
            for i in range(1, 11):
                attr = f'ent_worktime{i}'
                if hasattr(self, attr):
                    widget = getattr(self, attr)
                    if hasattr(widget, 'configure'):
                        try: 
                            widget['values'] = [''] + sorted_vals
                        except: pass
                grp_attr = f'worker_group{i}'
                if hasattr(self, grp_attr):
                    getattr(self, grp_attr).update_time_list(sorted_vals)
        elif config_key == 'equipments' or config_key == 'materials':
            # Both affect the unified material/equipment suggestion lists
            self.update_material_combo()
        elif config_key == 'test_items' or config_key == 'applied_codes':
            # These are handled by dynamic lambdas in _bind_combobox_word_suggest
            pass 
        elif config_key == 'warehouses':
            if hasattr(self, 'cb_warehouse'): self.cb_warehouse['values'] = sorted_vals
        elif config_key == 'vehicles':
            # 1. Update fixed panel
            if hasattr(self, 'fixed_vehicle_widget'):
                self.fixed_vehicle_widget.update_vehicle_list(sorted_vals)
            # 2. Update floating panels
            if hasattr(self, 'draggable_items'):
                for key, cont in self.draggable_items.items():
                    w = getattr(cont, '_widget', None)
                    if w and type(w).__name__ == 'VehicleInspectionWidget':
                        w.update_vehicle_list(sorted_vals)
            # (Legacy compatibility)
            if hasattr(self, 'vehicle_boxes'):
                for box in getattr(self, 'vehicle_boxes', []):
                    box.update_vehicle_list(sorted_vals)
            if hasattr(self, 'vehicle_inspections'):
                for key, widget_instance in getattr(self, 'vehicle_inspections', {}).items():
                    widget_instance.update_vehicle_list(sorted_vals)
        elif config_key == 'co_code' or config_key == '회사코드':
            # [NEW] Update company code combobox in daily usage
            if hasattr(self, 'cb_daily_co_code'): self.cb_daily_co_code['values'] = sorted_vals

        # Always save configuration after any list change
        self.save_tab_config()


        # 2. Update ALL draggable widgets (clones) that depend on this list
        for key, container in self.draggable_items.items():
            # Heuristic: if manage_list_key is missing but label suggests it's a worker/user/site
            m_key = getattr(container, '_manage_list_key', None)
            if not m_key:
                if hasattr(container, '_label_widget'):
                    lbl_text = container._label_widget.cget('text').lower()
                    if config_key == 'users' and any(x in lbl_text for x in ['작업자', '담당자', 'user', 'worker']):
                        m_key = 'users'
                        container._manage_list_key = 'users'
                    elif config_key == 'sites' and any(x in lbl_text for x in ['현장', 'site']):
                        m_key = 'sites'
                        container._manage_list_key = 'sites'
                    elif config_key == 'equipments' and any(x in lbl_text for x in ['장비', 'equip']):
                        m_key = 'equipments'
                        container._manage_list_key = 'equipments'
                    elif config_key == 'vehicles' and any(x in lbl_text for x in ['차량', 'vehicle']):
                        m_key = 'vehicles'
                        container._manage_list_key = 'vehicles'

            if m_key == config_key:
                if hasattr(container, '_widget') and hasattr(container._widget, 'config'):
                    try:
                        container._widget['values'] = sorted_vals
                    except:
                        pass

    def remove_box(self, key):
        """Intelligently remove a box: hide standard ones, destroy custom ones"""
        is_dynamic = key.startswith('memo_') or key.startswith('clone_') or key.startswith('checklist_') or key.startswith('vehicle_inspection_')
        
        # Allow deletion of dynamic widgets even if layout is locked
        if self.layout_locked and not is_dynamic: return
        
        if is_dynamic:
            # Permanent deletion for custom/dynamic items
            self.destroy_custom_widget(key)
        else:
            # Hiding for standard items (can be restored via Reset All)
            widget = self.draggable_items.get(key)
            if widget:
                # We reuse the existing hide_widget logic
                self.hide_widget(None, widget=widget)

    def destroy_custom_widget(self, key):
        """Destroy and remove from config any dynamic widget (clone or memo)"""
        # Allow destruction even if layout_locked (since it's a content management action)
        
        widget = self.draggable_items.get(key)
        if widget:
            widget.destroy()
            if key in self.draggable_items:
                del self.draggable_items[key]
            if key in self.memos:
                del self.memos[key]
            if key in self.checklists:
                del self.checklists[key]
            if key in self.vehicle_inspections:
                widget_to_remove = self.vehicle_inspections[key]
                if widget_to_remove in self.vehicle_boxes:
                    self.vehicle_boxes.remove(widget_to_remove)
                del self.vehicle_inspections[key]

            
            # Clean from in-memory config to prevent resurrection on next save_tab_config call
            if hasattr(self, 'tab_config') and 'draggable_geometries' in self.tab_config:
                if key in self.tab_config['draggable_geometries']:
                    del self.tab_config['draggable_geometries'][key]
            
            # Clean from config immediately to prevent resurrection on next load
            try:
                import json
                if os.path.exists(self.config_path):
                    with open(self.config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                    
                    if 'draggable_geometries' in config and key in config['draggable_geometries']:
                        del config['draggable_geometries'][key]
                        
                    with open(self.config_path, 'w', encoding='utf-8') as f:
                        json.dump(config, f, ensure_ascii=False, indent=2)
            except:
                pass
            
            self.save_tab_config()

    def rename_widget_label(self, key):
        """Show a simple dialog to rename a widget's label"""
        if self.layout_locked: return
        
        widget = self.draggable_items.get(key)
        if not widget: return
        
        current_text = ""
        if hasattr(widget, '_label_widget'):
            current_text = widget._label_widget.cget('text')
        elif key in self.memos:
            current_text = self.memos[key]['title_entry'].get()
        elif key in self.checklists:
            current_text = self.checklists[key]['title_entry'].get()
            
        new_name = simpledialog.askstring("이름 변경", "새 이름을 입력하세요:", initialvalue=current_text)
        if new_name is not None:
            if hasattr(widget, '_label_widget'):
                widget._label_widget.config(text=new_name)
            elif key in self.memos:
                self.memos[key]['title_entry'].delete(0, 'end')
                self.memos[key]['title_entry'].insert(0, new_name)
            elif key in self.checklists:
                self.checklists[key]['title_entry'].delete(0, 'end')
                self.checklists[key]['title_entry'].insert(0, new_name)
            self.save_tab_config()

    def clone_widget(self, key):
        """Create a clone of an existing widget as a new custom box"""
        if self.layout_locked: return
        
        orig = self.draggable_items.get(key)
        if not orig: return
        
        import time
        new_key = f"clone_{int(time.time() * 1000)}"
        
        label_text = ""
        if hasattr(orig, '_label_widget'):
            label_text = orig._label_widget.cget('text')
        
        if hasattr(orig, '_widget_class'):
            # It's a container created via create_draggable_container
            cont, w = self.create_draggable_container(
                self.entry_inner_frame, 
                label_text, 
                orig._widget_class, 
                new_key, 
                manage_list_key=getattr(orig, '_manage_list_key', None), # Pass manage_list_key
                **orig._widget_kwargs
            )
            
            # Copy value from original widget
            if hasattr(orig, '_widget'):
                try:
                    if hasattr(orig._widget, 'get'):
                        current_val = str(orig._widget.get()) # Ensure string
                        
                        # Try generic Entry-like setting (works for Entry and Combobox text area)
                        if hasattr(w, 'delete') and hasattr(w, 'insert'):
                            try:
                                w.delete(0, 'end')
                                w.insert(0, current_val)
                            except:
                                # Readonly comboboxes might fail delete/insert
                                pass
                                
                        # Try specific set method (Combobox, Scale, etc)
                        if hasattr(w, 'set'):
                            w.set(current_val)
                    elif hasattr(orig._widget, 'get_data') and hasattr(w, 'set_data'):
                        # For complex widgets like VehicleInspectionWidget
                        w.set_data(orig._widget.get_data())
                except Exception as e:
                    print(f"Failed to copy value: {e}")

            
            cont.place(x=50, y=50) # Start position
            self.save_tab_config()

        elif key in self.memos:
            # It's a memo
            content = self.memos[key]['text_widget'].get('1.0', 'end-1c')
            title = self.memos[key]['title_entry'].get()
            self.add_new_memo(initial_text=content, initial_title=title, key=new_key)
            self.save_tab_config()
        elif key in self.checklists:
            # It's a checklist
            self.duplicate_checklist(key)
            self.save_tab_config()

    def _bind_recursive(self, widget, target_container):
        """Recursively bind drag events to widget and its children"""
        from functools import partial
        
        # Bind events to this widget, targeting the container
        # Note: We use add=True to avoid overwriting existing bindings if possible, 
        # but for drag we usually want exclusive control or at least priority.
        # Here we just bind standard.
        
        # We need to capture the target_container in the callback
        widget.bind("<Button-3>", partial(self.on_drag_start, widget=target_container))
        widget.bind("<Shift-Button-3>", partial(self.on_resize_start, widget=target_container))
        widget.bind("<B3-Motion>", partial(self.on_mouse_motion, widget=target_container))
        widget.bind("<Double-Button-3>", partial(self.reset_widget_position, widget=target_container))
        widget.bind("<Control-Button-3>", partial(self.hide_widget, widget=target_container))
        widget.bind("<ButtonRelease-3>", partial(self.on_drag_stop, widget=target_container))
        
        # Recurse for children
        try:
            for child in widget.winfo_children():
                # Toplevel(DateEntry 달력 팝업 등) 내부는 절대 순회하지 않음
                # → winfo_children() 재귀만으로도 Toplevel이 화면에 나타날 수 있음
                if child.winfo_class() == 'Toplevel':
                    continue
                self._bind_recursive(child, target_container)
        except:
            pass

    def hide_widget(self, event, widget=None):
        """Hide a widget (Ctrl + Right Click)"""
        if self.layout_locked:
            return "break"
            
        if widget is None:
            widget = event.widget
            
        # [STABILITY FIX] Prevent hiding core widgets
        if hasattr(widget, '_config_key') and widget._config_key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
            messagebox.showwarning("보호됨", f"'{widget._config_key}' 항목은 화면의 필수 구성 요소이므로 숨길 수 없습니다.")
            return "break"
        
        # Remove from place layout
        if widget.winfo_manager() == 'place':
            widget.place_forget()
        elif widget.winfo_manager() == 'grid':
            widget.grid_forget()
            
        # Remove placeholder if exists
        self._remove_placeholder(widget)
        
        # Mark as hidden in config
        if hasattr(widget, '_config_key') and widget._config_key:
            try:
                import json
                if os.path.exists(self.config_path):
                    with open(self.config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                else:
                    config = {}
                
                if 'draggable_geometries' not in config:
                    config['draggable_geometries'] = {}
                
                if widget._config_key not in config['draggable_geometries']:
                    config['draggable_geometries'][widget._config_key] = {}
                
                config['draggable_geometries'][widget._config_key]['hidden'] = True
                
                with open(self.config_path, 'w', encoding='utf-8') as f:
                    json.dump(config, f, ensure_ascii=False, indent=2)
            except Exception as e:
                print(f"Failed to save hide status: {e}")
        return "break"

    def make_draggable(self, widget, config_key=None):
        """Make a widget draggable and resizable with right mouse button"""
        widget._config_key = config_key
        # Save original grid info for reset/placeholder
        # We need to do this immediately while it's still in the grid
        widget._original_grid_info = widget.grid_info()
        
        # Recursively bind to ensure clicking anywhere works
        self._bind_recursive(widget, widget)

    def make_header_draggable(self, widget, target_container):
        """Make a specific widget (header/label) draggable with Left Mouse Button targeting a container"""
        from functools import partial
        widget.bind("<Button-1>", partial(self.on_drag_start, widget=target_container))
        widget.bind("<B1-Motion>", partial(self.on_mouse_motion, widget=target_container))
        widget.bind("<ButtonRelease-1>", partial(self.on_drag_stop, widget=target_container))
        
    def reset_widget_position(self, event, widget=None):
        """Reset widget to original grid position"""
        if self.layout_locked:
            return "break"
            
        if widget is None:
            widget = event.widget
        
        # Remove from place layout
        widget.place_forget()
        
        # Remove placeholder if exists
        self._remove_placeholder(widget)
            
        # Restore to grid
        if hasattr(widget, '_original_grid_info'):
            widget.grid(**widget._original_grid_info)
        
        # Reset size variables if any
        if hasattr(widget, '_start_width'): del widget._start_width
        if hasattr(widget, '_start_height'): del widget._start_height
        
        # Remove from config
        if hasattr(widget, '_config_key') and widget._config_key:
            try:
                import json
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                
                if 'draggable_geometries' in config and widget._config_key in config['draggable_geometries']:
                    del config['draggable_geometries'][widget._config_key]
                
                # Backward compatibility: also check top level
                if widget._config_key in config:
                    del config[widget._config_key]
                    
                with open(self.config_path, 'w', encoding='utf-8') as f:
                    json.dump(config, f, ensure_ascii=False, indent=2)
                    
            except Exception as e:
                print(f"Failed to reset config: {e}")

    def reset_all_widgets_layout(self):
        """Reset all widgets to their original grid slots, unhide them, and reset paned window sashes"""
        if messagebox.askyesno("초기화", "모든 항목의 위치와 크기를 초기화하고 숨겨진 항목을 다시 표시하시겠습니까?\n(창 분할 위치도 초기화됩니다.)"):
            # 1. Reset draggable widgets
            for key, widget in list(self.draggable_items.items()):
                # Explicitly unhide if it was hidden
                if widget.winfo_manager() == '':
                    # For grid, we just grid it back
                    if hasattr(widget, '_original_grid_info'):
                         widget.grid(**widget._original_grid_info)
                
                self.reset_widget_position(None, widget=widget)
            
            # 1.05 [NEW] Use custom defaults if they exist
            if os.path.exists(self.config_path):
                try:
                    with open(self.config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                    
                    custom_defaults = config.get('custom_default_geometries', {})
                    if custom_defaults:
                        # Clear existing geometries to ensure fresh start
                        if 'draggable_geometries' not in config: config['draggable_geometries'] = {}
                        
                        for key, geo in custom_defaults.items():
                            if key in self.draggable_items:
                                w = self.draggable_items[key]
                                w.place(x=geo['x'], y=geo['y'], width=geo['width'], height=geo['height'])
                                # Also update config so it persists as the new current layout
                                config['draggable_geometries'][key] = geo
                        
                        # Save the updated config back to disk and memory
                        with open(self.config_path, 'w', encoding='utf-8') as f:
                            json.dump(config, f, ensure_ascii=False, indent=2)
                        self.tab_config = config
                except Exception as e:
                    print(f"Error applying custom defaults: {e}")
                
            # 1.1 Restore parent propagation safely and reset size
            if hasattr(self, 'entry_inner_frame'):
                self.entry_inner_frame.pack_propagate(True)
                self.entry_inner_frame.grid_propagate(True)
                
                # Perform a layout update pass
                self.root.update_idletasks()
                
                # Refresh scrollregion if in canvas
                if hasattr(self, 'entry_canvas'):
                    self.entry_canvas.configure(scrollregion=self.entry_canvas.bbox("all"))
            
            # 2. Reset sash positions (splitters)
            try:
                if hasattr(self, 'daily_usage_paned'):
                    # Give more space to the entry form by default (500px)
                    getattr(self.daily_usage_paned, "sashpos", lambda *args: 500)(0, 500) 
                if hasattr(self, 'daily_history_paned'):
                    total_w = self.daily_history_paned.winfo_width()
                    if total_w > 100:
                        self.daily_history_paned.sashpos(0, int(total_w * 0.7))
                    else:
                        self.daily_history_paned.sashpos(0, 800)
            except:
                pass

            self.save_tab_config()
            messagebox.showinfo("완료", "레이아웃과 창 분할 위치가 초기화되었습니다.")

    def save_current_layout_as_default(self):
        """Save the current layout as the custom 'reset' default"""
        try:
            current_geos = {}
            for key, widget in self.draggable_items.items():
                # [FIX] Capture all managed widgets (grid or place) 
                # This ensures the snapshot is complete even if some items weren't moved.
                manager = widget.winfo_manager()
                if manager in ['place', 'grid']:
                    current_geos[key] = {
                        'x': widget.winfo_x(),
                        'y': widget.winfo_y(),
                        'width': widget.winfo_width(),
                        'height': widget.winfo_height(),
                        'hidden': False
                    }
            
            if not current_geos:
                messagebox.showwarning("실패", "수동으로 배치된(드래그된) 항목이 없어 기본값으로 저장할 수 없습니다.")
                return

            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            else:
                config = {}
                
            config['custom_default_geometries'] = current_geos
            
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
                
            messagebox.showinfo("완료", "현재 레이아웃이 '기본 배치'로 저장되었습니다.\n앞으로 [초기화] 시 이 배치로 되돌아갑니다.")
        except Exception as e:
            messagebox.showerror("오류", f"기본 배치 저장 중 오류 발생: {e}")
            
    def toggle_layout_lock(self):
        """Toggle layout locking state"""
        self.layout_locked = not self.layout_locked
        if hasattr(self, 'btn_lock_layout'):
            if self.layout_locked:
                self.btn_lock_layout.config(text="🔒 배치 고정됨")
                self.style.configure("Lock.TButton", foreground="black")
            else:
                self.btn_lock_layout.config(text="🔓 배치 수정 중")
                self.style.configure("Lock.TButton", foreground="red")
        
        # Save layout lock state to config immediately
        if not hasattr(self, 'tab_config'):
            self.tab_config = {}
        self.tab_config['layout_locked'] = self.layout_locked
        
        # [FIX] Unlocking is no longer destructive. 
        # It simply enables movement without resetting positions.
        if self.layout_locked:
            print("Layout locked - current positions preserved")
        else:
            print("Layout unlocked - movement enabled")
        
        # Force save the state immediately to file
        self.save_tab_config(force=True)
        
        print(f"Layout lock {'enabled' if self.layout_locked else 'disabled'}")

    def on_drag_stop(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import on_drag_stop_impl
        return on_drag_stop_impl(self, *args, **kwargs)


    def add_new_memo(self, initial_text="", initial_title="메모", key=None):
        """Create a new movable/editable/copyable memo box with editable title"""
        import time
        if key is None:
            key = f"memo_{int(time.time() * 1000)}"
        
        # Create container

        memo_container = ttk.LabelFrame(self.entry_inner_frame)
        
        # Controls frame (top of memo)
        ctrl_frame = ttk.Frame(memo_container)
        ctrl_frame.pack(fill='x', side='top')
        
        # Editable Title Entry
        title_entry = ttk.Entry(ctrl_frame, font=('Arial', 9, 'bold'), width=15)
        title_entry.pack(side='left', padx=2)
        title_entry.insert(0, initial_title)
        title_entry.bind('<FocusOut>', lambda e: self.save_tab_config())
        
        btn_copy = ttk.Button(ctrl_frame, text="📋", width=3, command=lambda: self.duplicate_memo(key))
        btn_copy.pack(side='right')
        
        btn_del = ttk.Button(ctrl_frame, text="❌", width=3, command=lambda: self.remove_box(key))
        btn_del.pack(side='right')
        
        # Text area
        text_area = tk.Text(memo_container, wrap='word', height=5, width=30, font=('Arial', 10), bg=self.theme_bg, highlightthickness=0)
        text_area.pack(fill='both', expand=True, padx=2, pady=2)
        text_area.insert('1.0', initial_text)
        
        # Bind text change to auto-save
        text_area.bind('<FocusOut>', lambda e: self.save_tab_config())
        
        # Make draggable (Right Click)
        self.make_draggable(memo_container, key)
        
        # Make header draggable (Left Click)
        self.make_header_draggable(ctrl_frame, memo_container)
        self.draggable_items[key] = memo_container
        self.memos[key] = {
            'container': memo_container, 
            'text_widget': text_area,
            'title_entry': title_entry
        }
        
        # Initial placement if not loaded from config (will be handled by load_tab_config if exists)
        if key not in getattr(self, '_loading_memos', []):
            memo_container.place(x=1450, y=50) # Default start position
        
        return memo_container

    def duplicate_memo(self, key):
        """Duplicate an existing memo with its title and content"""
        if self.layout_locked: return
        if key in self.memos:
            content = self.memos[key]['text_widget'].get('1.0', 'end-1c')
            title = self.memos[key]['title_entry'].get()
            self.add_new_memo(initial_text=content, initial_title=title)
            self.save_tab_config()

    def add_new_checklist(self, initial_data=None, initial_title="체크리스트", key=None):
        """Create a new movable/editable checklist box"""
        import time
        if key is None:
            key = f"checklist_{int(time.time() * 1000)}"
        
        # Create container
        check_container = ttk.LabelFrame(self.entry_inner_frame)
        
        # Controls frame (top of checklist)
        ctrl_frame = ttk.Frame(check_container)
        ctrl_frame.pack(fill='x', side='top')
        
        # Editable Title Entry
        title_entry = ttk.Entry(ctrl_frame, font=('Arial', 9, 'bold'), width=15)
        title_entry.pack(side='left', padx=2)
        title_entry.insert(0, initial_title)
        title_entry.bind('<FocusOut>', lambda e: self.save_tab_config())
        
        btn_copy = ttk.Button(ctrl_frame, text="📋", width=3, command=lambda: self.duplicate_checklist(key))
        btn_copy.pack(side='right')
        
        btn_del = ttk.Button(ctrl_frame, text="❌", width=3, command=lambda: self.remove_box(key))
        btn_del.pack(side='right')
        
        # Add Item Area
        add_frame = ttk.Frame(check_container)
        add_frame.pack(fill='x', padx=2, pady=2)
        
        new_item_var = tk.StringVar()
        entry_new = ttk.Entry(add_frame, textvariable=new_item_var, width=20)
        entry_new.pack(side='left', fill='x', expand=True)
        
        def add_item(event=None):
             text = new_item_var.get().strip()
             if text:
                 self.add_checklist_item(item_frame, text, False, key)
                 new_item_var.set("")
                 self.save_tab_config()
        
        def _adjust_parent_height(self, parent, force=False):
            """Adjust parent frame height with performance check"""
            try:
                # Only update idletasks if forced or we're not in the middle of a high-speed interaction
                # This is the single biggest cause of UI stutter.
                if force:
                    parent.update_idletasks()
            except:
                pass
        
        entry_new.bind('<Return>', add_item)
        btn_add = ttk.Button(add_frame, text="➕", width=3, command=add_item)
        btn_add.pack(side='right')

        # Scrollable Frame for Items
        canvas_frame = ttk.Frame(check_container)
        canvas_frame.pack(fill='both', expand=True, padx=2, pady=2)
        
        canvas = tk.Canvas(canvas_frame, height=100, width=200, bg=self.theme_bg, highlightthickness=0) # Match theme
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        item_frame = ttk.Frame(canvas)
        
        item_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=item_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Scroll binding is handled globally in MaterialManager.__init__

        # Make draggable
        self.make_draggable(check_container, key)
        self.draggable_items[key] = check_container
        self.checklists[key] = {
            'container': check_container,
            'title_entry': title_entry,
            'item_frame': item_frame,
            'items': [] # List of item widgets/vars will be managed dynamically via children of item_frame
        }
        
        # Add initial items if provided
        if initial_data:
            for item in initial_data:
                self.add_checklist_item(item_frame, item.get('text', ''), item.get('checked', False), key)
        
        # Initial placement
        if key not in getattr(self, '_loading_memos', []): # Reuse loading flag or logic
            check_container.place(x=1450, y=50)
            
        return check_container

    def save_single_vehicle_data(self, widget):
        """Save ONLY the requested vehicle's data as a standalone entry in the daily log"""
        try:
            # 1. Collect minimal context
            usage_date = str(self.ent_daily_date.get_date())
            site = self.cb_daily_site.get().strip()
            if not site:
                messagebox.showwarning("입력 필요", "차량 점검을 기록할 '현장'을 선택해주세요.")
                return

            # 2. Collect vehicle data
            v_data = widget.get_data()
            v_no = v_data.get('vehicle_info', '').strip()
            if not v_no:
                messagebox.showwarning("입력 필요", "차량번호를 입력해주세요.")
                return

            # Confirm with user
            if not messagebox.askyesno("확인", f"[{v_no}] 차량 점검 정보만 단독으로 저장하시겠습니까?"):
                return

            # 3. Create record (Everything else is zero/empty)
            reserved = ['vehicle_info', 'mileage', 'remarks', '_raw_mileage']
            checks = "|".join([f"{k}:{v}" for k, v in v_data.items() if k not in reserved and v])

            record = {
                'Date': usage_date,
                '업체명': self.cb_daily_company.get().strip() or "현장기록",
                'Site': site,
                'User': "차량점검", # Visual tag in worker column
                'WorkTime': "",
                '장비명': self.cb_daily_equip.get().strip(),
                '검사방법': self.cb_daily_test_method.get().strip(),
                '회사코드': "",
                'Usage': 0.0, '단가': 0.0, '출장비': 0.0, '검사비': 0.0,
                '검사량': 0.0,
                'OT': 0.0, 'OT금액': 0.0,
                'MaterialID': "차량점검",
                '차량번호': v_no,
                '주행거리': v_data.get('_raw_mileage', '0'),
                '차량점검': checks,
                '차량비고': v_data.get('remarks', ''),
                'EntryTime': datetime.datetime.now(),
                '(Full작업자)': "차량점검"
            }
            # Initialize RTK/NDT columns as 0
            for rtk_cat in ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타", "RTK총계"]:
                record[f'RTK_{rtk_cat}'] = 0
            for ndt_cat in ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]:
                record[f'NDT_{ndt_cat}'] = 0

            # 4. Append and Save
            new_df = pd.DataFrame([record])
            self.daily_usage_df = pd.concat([self.daily_usage_df, new_df], ignore_index=True)
            self.save_data()
            
            # 5. UI Updates
            self.update_daily_usage_view()
            self.refresh_inquiry_filters()
            
            messagebox.showinfo("성공", f"[{v_no}] 차량 정보가 기록에 추가되었습니다.\n(실적 없이 차량 단독 기록으로 저장됨)")
            
            # [V12_QUICK_MULTI_VEHICLE] Reset only vehicle-specific fields for next entry
            if hasattr(widget, 'reset_fields'):
                widget.reset_fields()
            
        except Exception as e:
            messagebox.showerror("오류", f"차량 정보 저장 중 오류 발생: {e}")

    def add_vehicle_inspection_box(self, initial_data=None, key=None):
        """Create a new movable vehicle inspection box"""
        import datetime
        if key is None:
            key = f"vehicle_inspection_{int(datetime.datetime.now().timestamp() * 1000)}"
            
        # Create draggable container
        container, widget = self.create_draggable_container(
            self.entry_inner_frame, "차량 점검", VehicleInspectionWidget, key,
            manage_list_key='vehicles',
            theme_bg=getattr(self, 'theme_bg', '#f0f0f0'),
            vehicle_list=getattr(self, 'vehicles', []),
            on_save=self.save_single_vehicle_data
        )
        
        # Track specifically for data gathering
        self.vehicle_inspections[key] = widget
        if widget not in self.vehicle_boxes:
            self.vehicle_boxes.append(widget)
        
        # Load initial data if provided
        if initial_data:
            widget.set_data(initial_data)
            
        # Initial placement if new (Below the trigger button in the main entry frame)
        if key not in getattr(self, '_loading_memos', []):
            container.place(x=1450, y=50)
            
        return widget

    def add_checklist_item(self, parent_frame, text, checked, checklist_key):
        """Add a single item row to the checklist"""
        row_frame = ttk.Frame(parent_frame)
        row_frame.pack(fill='x', pady=1)
        
        var = tk.BooleanVar(value=checked)
        cb = ttk.Checkbutton(row_frame, variable=var, command=lambda: self.save_tab_config())
        cb.pack(side='left')
        
        entry = ttk.Entry(row_frame)
        entry.insert(0, text)
        entry.pack(side='left', fill='x', expand=True)
        entry.bind('<FocusOut>', lambda e: self.save_tab_config())
        
        def delete_this_item():
            row_frame.destroy()
            self.save_tab_config()
            
        btn_del_item = ttk.Label(row_frame, text="❌", font=('Arial', 8), cursor='hand2', foreground='gray')
        btn_del_item.pack(side='right', padx=2)
        btn_del_item.bind('<Button-1>', lambda e: delete_this_item())
        
        # Store refs in widget for retrieval during save
        row_frame._checklist_data = {'var': var, 'entry': entry}

    def duplicate_checklist(self, key):
        """Duplicate an existing checklist"""
        if self.layout_locked: return
        if key in self.checklists:
            # get current data
            original_items = []
            item_frame = self.checklists[key]['item_frame']
            for child in item_frame.winfo_children():
                if hasattr(child, '_checklist_data'):
                    data = child._checklist_data
                    original_items.append({
                        'text': data['entry'].get(),
                        'checked': data['var'].get()
                    })
            
            title = self.checklists[key]['title_entry'].get()
            self.add_new_checklist(initial_data=original_items, initial_title=title)
            self.save_tab_config()


    def on_drag_start(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import on_drag_start_impl
        return on_drag_start_impl(self, *args, **kwargs)
        
    def on_resize_start(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import on_resize_start_impl
        return on_resize_start_impl(self, *args, **kwargs)
        
    def on_mouse_motion(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import on_mouse_motion_impl
        return on_mouse_motion_impl(self, *args, **kwargs)

    def _update_widget_position(self, event, widget):
        """Internal helper to calculate and set widget position/size during interaction"""
        dx = event.x_root - widget._drag_start_root_x
        dy = event.y_root - widget._drag_start_root_y
        parent = widget.master
        parent_w = parent.winfo_width()
        parent_h = parent.winfo_height()
        
        if widget._interaction_mode == 'move':
            new_x = widget._drag_start_pos_x + dx
            new_y = widget._drag_start_pos_y + dy
            widget_w = widget.winfo_width()
            widget_h = widget.winfo_height()
            
            # Clamp
            if new_x < 0: new_x = 0
            elif new_x + widget_w > parent_w: new_x = max(0, parent_w - widget_w)
            
            # Loosen Y clamping: allow moving slightly below parent to trigger growth
            if new_y < 0: new_y = 0
            # Instead of strict clamp at parent_h, allow a small overflow to trigger resize logic
            elif new_y + widget_h > parent_h + 100: 
                new_y = parent_h + 100 - widget_h

            
            if widget.winfo_manager() != 'place':
                # Save original grid info and explicitly un-grid
                if not hasattr(widget, '_original_grid_info') and widget.grid_info():
                    widget._original_grid_info = widget.grid_info()
                self._ensure_placeholder(widget)
                widget.grid_forget() # Explicitly un-grid before placing
                widget.place(width=widget_w, height=widget_h)
                
            widget.place(x=new_x, y=new_y)
            widget.lift()
            
            # Auto-update scroll region if we're dragging near bounds
            self._ensure_canvas_scroll_region()
            
        elif widget._interaction_mode == 'resize':
            new_width = max(50, widget._start_width + dx)
            new_height = max(20, widget._start_height + dy)
            current_x = widget.winfo_x()
            current_y = widget.winfo_y()
            
            # Clamp
            if current_x + new_width > parent_w: new_width = max(50, parent_w - current_x)
            if current_y + new_height > parent_h: new_height = max(20, parent_h - current_y)
            
            if widget.winfo_manager() != 'place':
                # Save original grid info and explicitly un-grid
                if not hasattr(widget, '_original_grid_info') and widget.grid_info():
                    widget._original_grid_info = widget.grid_info()
                self._ensure_placeholder(widget)
                widget.grid_forget() # Explicitly un-grid before placing
                widget.place(x=current_x, y=current_y)
                
            widget.place(width=new_width, height=new_height)
            widget.lift()
        
        return "break"
        
    def _apply_push_down_logic(self, dragged_widget):
        """Recursively push widgets down if they overlap with the dragged widget"""
        try:
            # Current bounding box of the dragging widget
            x1 = dragged_widget.winfo_x()
            y1 = dragged_widget.winfo_y()
            w1 = dragged_widget.winfo_width()
            h1 = dragged_widget.winfo_height()
            
            padding = 10
            
            # Parent of the widgets
            parent = dragged_widget.master
            
            # OPTIMIZATION: Early filter candidates by parent to avoid massive iterations
            candidates = [w for w in self.draggable_items.values() if w.master == parent]
            
            for other in candidates:
                if other == dragged_widget:
                    continue
                
                # [STABILITY FIX] Skip if the 'other' widget's BOTTOM is ABOVE the dragged widget's TOP.
                # Strictly ignore anything physically above the current interaction zone.
                if (other.winfo_y() + other.winfo_height()) <= y1:
                    continue

                # Skip if not in the same parent
                if other.master != parent:
                    continue
                
                # Skip if hidden
                if not other.winfo_ismapped():
                    continue
                
                # Get current pos/size of the 'other' widget
                x2 = other.winfo_x()
                y2 = other.winfo_y()
                w2 = other.winfo_width()
                h2 = other.winfo_height()
                
                # 1. Horizontal Overlap Check
                if (x1 < x2 + w2) and (x1 + w1 > x2):
                    # 2. Vertical Collision Check
                    # If dragged_widget is hitting 'other' from the top
                    # We check if the bottom of dragged_widget is below the top of other
                    if y1 < y2 and (y1 + h1) > y2:
                        # New target Y for 'other'
                        new_y2 = y1 + h1 + padding
                        
                        # Only move if it's actually pushing it DOWN
                        if y2 < new_y2:
                            if other.winfo_manager() == 'grid':
                                # Convert grid to place
                                self._ensure_placeholder(other)
                                other.lift()
                                other.place(x=x2, y=new_y2, width=w2, height=h2)
                            else:
                                # Just update place
                                other.place(y=new_y2)
                            
                            # 3. Recursive Push
                            self._apply_push_down_logic(other)
        except Exception as e:
            # Silent fail to avoid interrupting drag
            print(f"Error in push-down logic: {e}")

    def _adjust_parent_height(self, parent, force=False):
        """Adjust parent frame height with performance check"""
        try:
            # Only update idletasks if forced or we're not in the middle of a high-speed interaction
            # This is the single biggest cause of UI stutter.
            if force:
                parent.update_idletasks()
            # 2. Start with the bounding box of all GRIDDED items
            try:
                # grid_bbox returns (x, y, width, height) of the grid
                bbox = parent.grid_bbox()
                required_h = bbox[1] + bbox[3] if bbox[3] > 0 else 0
                required_w = bbox[0] + bbox[2] if bbox[2] > 0 else 0
            except:
                required_h = 0
                required_w = 0

            # 3. Handle PLACE items as well (dragged/custom items)
            for child in parent.winfo_children():
                try:
                    manager = child.winfo_manager()
                    if manager == 'place':
                        info = child.place_info()
                        # Get relative or absolute y + height
                        y = child.winfo_y()
                        h = child.winfo_height()
                        x = child.winfo_x()
                        w = child.winfo_width()
                        required_h = max(required_h, y + h)
                        required_w = max(required_w, x + w)
                except:
                    continue
            
            # Add some padding
            new_height = required_h + 30
            
            # [STABILITY FIX] Guard against collapse but allow growth
            # Width is now dictated by canvas parent in responsive mode.
            if new_height < 500: new_height = 800
            
            # Only resize if the required dimensions are significantly different
            current_h = parent.winfo_height()
            if abs(current_h - new_height) > 10:
                 parent.config(height=new_height)
                 
                 # [RECURSIVE FIX] If this parent has a master that is part of the entry system,
                 # adjust its height too. 
                 if parent.master and parent.master != self.entry_inner_frame.master: # Don't go above canvas
                     self._adjust_parent_height(parent.master, force=force)
                     
        except Exception as e:
            print(f"Error adjusting parent height: {e}")


    def _ensure_placeholder(self, widget, width=None, height=None):
        """Ensure a placeholder exists in the grid where the widget used to be"""
        # Map widget to a placeholder attribute name dynamically
        # We can use the widget id or a dictionary, but simpler to attach to widget
        if not hasattr(widget, '_placeholder'):
            # Use provided dims or current widget dims
            w = width if width is not None else widget.winfo_width()
            h = height if height is not None else widget.winfo_height()
            
            # Create a frame to hold the space
            widget._placeholder = ttk.Frame(widget.master, width=w, height=h)
            
            # Grid it at the original position
            if hasattr(widget, '_original_grid_info'):
                widget._placeholder.grid(**widget._original_grid_info)
                # Ensure the placeholder doesn't shrink and holds its size
                widget._placeholder.grid_propagate(False)
                
    def _remove_placeholder(self, widget):
        """Remove placeholder for a widget"""
        if hasattr(widget, '_placeholder'):
            widget._placeholder.destroy()
            del widget._placeholder

        

    def setup_daily_usage_tab(self, *args, **kwargs):
        from site_apps.central.src.views.daily_usage_view import setup_daily_usage_tab_impl
        return setup_daily_usage_tab_impl(self, *args, **kwargs)

    def on_recent_record_click(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import on_recent_record_click_impl
        return on_recent_record_click_impl(self, *args, **kwargs)

    def delete_recent_entry(self, event=None):
        """최근 기록 테이블에서 선택한 항목을 삭제"""
        selection = self.tv_recent.selection()
        if not selection:
            return
            
        import tkinter.messagebox as messagebox
        result = messagebox.askyesno("삭제 확인", "선택한 최근 기록을 삭제하시겠습니까?\n(삭제 시 차감되었던 재고도 자동으로 환원됩니다.)")
        if not result:
            return
            
        indices_to_delete = []
        for item in selection:
            tags = self.tv_recent.item(item, 'tags')
            if tags:
                try:
                    df_idx = int(tags[0])
                    indices_to_delete.append(df_idx)
                except ValueError:
                    continue
                    
        if not indices_to_delete:
            return
            
        try:
            valid_indices_to_delete = []
            for idx in indices_to_delete:
                if idx not in self.daily_usage_df.index:
                    continue
                
                valid_indices_to_delete.append(idx)
                entry = self.daily_usage_df.loc[idx]
                site = entry.get('Site', '')
                usage_date = pd.to_datetime(entry.get('Date'))
                
                if not self.transactions_df.empty:
                    trans_mask = (
                        (pd.to_datetime(self.transactions_df['Date'], errors='coerce').dt.normalize() == pd.to_datetime(usage_date).normalize()) &
                        (self.transactions_df['Site'].astype(str) == str(site)) &
                        (self.transactions_df['Type'] == 'OUT') &
                        (self.transactions_df['Note'].str.contains(f"{site} 현장 사용", na=False, regex=False))
                    )
                    self.transactions_df = self.transactions_df[~trans_mask]
                    
            if valid_indices_to_delete:
                self.daily_usage_df = self.daily_usage_df.drop(valid_indices_to_delete)
                self.daily_usage_df = self.daily_usage_df.reset_index(drop=True)
            
            if self.save_data():
                # messagebox.showinfo("삭제 완료", "기록이 성공적으로 삭제되었습니다.")
                self.update_daily_usage_view()
                self.update_recent_entries_view() # Ensure the mini table is updated!
                self.update_stock_view()
                self.update_transaction_view()
                self.refresh_inquiry_filters()
                
                # [FIX] Automatically refresh the Site tab and Query tab
                if hasattr(self, 'budget_view_tree'):
                    self.update_budget_site_view()
                if hasattr(self, 'query_tree') and hasattr(self, 'cb_filter_year'):
                    try:
                        y = int(self.cb_filter_year.get().replace('년', ''))
                        m = int(self.cb_filter_month.get().replace('월', ''))
                        self.update_monthly_usage_view(y, m)
                    except:
                        pass
        except Exception as e:
            messagebox.showerror("삭제 오류", f"기록 삭제 중 오류가 발생했습니다: {e}")

    def update_recent_entries_view(self):
        """오늘 입력된 내역을 미니 테이블에 업데이트"""
        if not hasattr(self, 'tv_recent') or getattr(self, 'daily_usage_df', None) is None:
            return
            
        for item in self.tv_recent.get_children():
            self.tv_recent.delete(item)
            
        if self.daily_usage_df.empty:
            return
            
        try:
            # 오늘 작업일자 기록을 우선 표시하고, 없으면 전체 최근 기록을 표시한다.
            today = datetime.datetime.now().date()
            work_dates = pd.to_datetime(self.daily_usage_df['Date'], errors='coerce')
            recent_df = self.daily_usage_df[work_dates.dt.date == today].copy()
            
            # 선택/수정에서 원본 행 번호를 사용하므로 DataFrame 인덱스는 유지한다.
            if recent_df.empty:
                recent_df = self.daily_usage_df.copy()

            recent_df['_recent_date'] = pd.to_datetime(recent_df['Date'], errors='coerce')
            if 'EntryTime' in recent_df.columns:
                recent_df['_recent_entry_time'] = pd.to_datetime(
                    recent_df['EntryTime'], errors='coerce'
                )
                recent_df = recent_df.sort_values(
                    by=['_recent_date', '_recent_entry_time'],
                    ascending=[False, False],
                    na_position='last',
                )
            else:
                recent_df = recent_df.sort_values(
                    by='_recent_date', ascending=False, na_position='last'
                )
            recent_df = recent_df.head(30)
                
            for idx, row in recent_df.iterrows():
                # Extract first worker
                first_worker = str(row.get('User', '')).strip()
                if first_worker.lower() == 'nan':
                    first_worker = ''
                elif first_worker:
                    import re
                    # Remove shift markers like (주간), (야간) if present
                    match = re.match(r"\((주간|야간|휴일|주야간)\)\s*(.*)", first_worker)
                    if match:
                        first_worker = match.group(2).strip()
                        
                date_str = str(row.get('Date', '')).split(' ')[0]
                mat_id = row.get('MaterialID', '')
                mat_name = self.get_material_display_name(mat_id) if hasattr(self, 'get_material_display_name') else mat_id
                
                # [FIX] If Material is empty but Equipment is present, show Equipment in the TreeView
                if not mat_name or str(mat_name).strip() == '':
                    equip_name = str(row.get('장비명', '')).strip()
                    if equip_name and equip_name.lower() != 'nan':
                        mat_name = f"[장비] {equip_name}"
                
                loc_type = str(row.get('구분', '')).strip()
                if not loc_type or loc_type.lower() == 'nan':
                    site = str(row.get('Site', '')).strip()
                    item = str(row.get('검사품명', '')).strip().upper()
                    if '관리소' in site or '관리소' in item or 'STATION' in item or 'V/S' in item or 'B/V' in item:
                        loc_type = '플랜트(관리소)'
                    else:
                        loc_type = '열배관'
                
                values = (
                    idx,
                    date_str,
                    row.get('Site', ''),
                    loc_type,
                    row.get('검사방법', ''),
                    row.get('검사품명', ''),
                    mat_name,
                    row.get('Usage', ''),
                    first_worker,
                    row.get('검사구분', ''),
                    row.get('조인트수', ''),
                    row.get('불량수', ''),
                    row.get('관경(Inch)', '')
                )
                self.tv_recent.insert('', 'end', values=values, tags=(str(idx),))
                
            # Scroll to bottom
            if self.tv_recent.get_children():
                last_item = self.tv_recent.get_children()[-1]
                self.tv_recent.see(last_item)
        except Exception as e:
            print(f"DEBUG: Error updating recent entries view: {e}")

    def setup_daily_usage_query_tab(self):
        """Setup the daily usage query tab"""
        display_frame = ttk.Frame(self.tab_daily_usage_query)
        display_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # KPI Summary Frame
        self.kpi_frame = ttk.LabelFrame(display_frame, text="조회 기간 요약 (KPI)")
        self.kpi_frame.pack(fill='x', padx=5, pady=(0, 5))
        self.lbl_kpi_summary = ttk.Label(self.kpi_frame, text="데이터를 불러오는 중...", font=('Malgun Gothic', 10, 'bold'), foreground='#00529B')
        self.lbl_kpi_summary.pack(side='left', padx=10, pady=5)
        
        # Filter controls
        filter_frame = ttk.Frame(display_frame)
        filter_frame.pack(fill='x', padx=5, pady=0)
        
        # --- Row 1: Date Filters ---
        date_row = ttk.Frame(filter_frame)
        date_row.pack(fill='x', pady=2)
        
        ttk.Label(date_row, text="시작일:").pack(side='left', padx=5)
        self.ent_daily_start_date = DateEntry(date_row, width=12, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        self.ent_daily_start_date.pack(side='left', padx=5)
        # Default to showing all history (starting from 2024)
        start_date = datetime.datetime(2024, 1, 1)
        self.ent_daily_start_date.set_date(start_date)
        
        ttk.Label(date_row, text="종료일:").pack(side='left', padx=5)
        self.ent_daily_end_date = DateEntry(date_row, width=12, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        self.ent_daily_end_date.pack(side='left', padx=5)
        self.ent_daily_end_date.set_date(datetime.datetime.now())
        
        # --- Row 2: Search Filters ---
        filter_row = ttk.Frame(filter_frame)
        filter_row.pack(fill='x', pady=2)
        
        ttk.Label(filter_row, text="업체명:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_company = ttk.Combobox(filter_row, width=12)
        self.cb_daily_filter_company.pack(side='left', padx=2)
        self.cb_daily_filter_company.set('전체')
        tk.Button(filter_row, text="⚙️", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                  command=lambda: self.open_list_management_dialog('companies', target_cb=self.cb_daily_filter_company)).pack(side='left', padx=(0, 5))
        
        ttk.Label(filter_row, text="현장:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_site = ttk.Combobox(filter_row, width=12)
        self.cb_daily_filter_site.pack(side='left', padx=2)
        self.cb_daily_filter_site.set('전체')
        tk.Button(filter_row, text="⚙️", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                  command=lambda: self.open_list_management_dialog('sites', target_cb=self.cb_daily_filter_site)).pack(side='left', padx=(0, 5))
        
        ttk.Label(filter_row, text="품목명:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_material = ttk.Combobox(filter_row, width=15)
        self.cb_daily_filter_material.pack(side='left', padx=2)
        self.cb_daily_filter_material.set('전체')
        tk.Button(filter_row, text="⚙️", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                  command=lambda: self.open_list_management_dialog('materials', target_cb=self.cb_daily_filter_material)).pack(side='left', padx=(0, 5))
        
        ttk.Label(filter_row, text="장비명:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_equipment = ttk.Combobox(filter_row, width=15)
        self.cb_daily_filter_equipment.pack(side='left', padx=2)
        self.cb_daily_filter_equipment.set('전체')
        tk.Button(filter_row, text="⚙️", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                  command=lambda: self.open_list_management_dialog('equipments', target_cb=self.cb_daily_filter_equipment)).pack(side='left', padx=(0, 5))
        
        ttk.Label(filter_row, text="작업자:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_worker = ttk.Combobox(filter_row, width=10)
        self.cb_daily_filter_worker.pack(side='left', padx=2)
        self.cb_daily_filter_worker.set('전체')
        
        ttk.Label(filter_row, text="분류:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_shift = ttk.Combobox(filter_row, width=8, state="readonly", values=["전체", "주간", "야간", "주야간", "휴일"])
        self.cb_daily_filter_shift.pack(side='left', padx=2)
        self.cb_daily_filter_shift.set('전체')
        
        # --- Row 3: Action Buttons (Top) ---
        btn_row = ttk.Frame(filter_frame)
        btn_row.pack(fill='x', pady=2)
        
        # --- Row 4: Action Buttons (Bottom) ---
        btn_row2 = ttk.Frame(filter_frame)
        btn_row2.pack(fill='x', pady=2)
        
        btn_filter = ttk.Button(btn_row, text="조회", style='Action.TButton', command=self.update_daily_usage_view)
        btn_filter.pack(side='left', padx=5)
        
        btn_filter_reset = ttk.Button(btn_row, text="♻️ 필터 초기화", command=self.reset_daily_usage_filters)
        btn_filter_reset.pack(side='left', padx=5)
        
        btn_delete = ttk.Button(btn_row, text="선택 항목 삭제", command=self.delete_daily_usage_entry)
        btn_delete.pack(side='left', padx=10)
        
        btn_edit = ttk.Button(btn_row, text="선택 항목 수정", command=self.open_edit_daily_usage_dialog)
        btn_edit.pack(side='left', padx=5)
        
        # Move export/report related buttons to btn_row2 to avoid horizontal cutoff
        btn_export = ttk.Button(btn_row2, text="엑셀 내보내기", command=self.export_daily_usage_history)
        btn_export.pack(side='left', padx=5)
        
        btn_export_invoice = ttk.Button(btn_row2, text="기성청구서 출력", command=self.export_invoice_excel)
        btn_export_invoice.pack(side='left', padx=5)
        
        btn_export_all = ttk.Button(btn_row2, text="전체 기록 내보내기", command=self.export_all_daily_usage)
        btn_export_all.pack(side='left', padx=5)
        
        btn_col_manage = ttk.Button(btn_row2, text="컬럼 관리", command=self.show_column_visibility_dialog)
        btn_col_manage.pack(side='left', padx=10)

        btn_ndt_report = ttk.Button(btn_row2, text="📊 진도보고서 출력", command=self.export_monthly_ndt_report)
        btn_ndt_report.pack(side='left', padx=5)

        btn_weekly_report = ttk.Button(btn_row2, text="🗓️ 주간 업무보고 출력", command=self.export_weekly_report)
        btn_weekly_report.pack(side='left', padx=5)

        # Dedicated Save Button for the List View
        self.btn_daily_save_list = ttk.Button(btn_row2, text="💾 변경사항 저장", command=self.save_all_daily_usage_changes, style='Accent.TButton' if 'Accent.TButton' in self.style.theme_names() else 'TButton')
        self.btn_daily_save_list.pack(side='left', padx=10)

        # Bindings
        filter_widgets = [
            self.cb_daily_filter_site, self.cb_daily_filter_company, self.cb_daily_filter_material, 
            self.cb_daily_filter_equipment, self.cb_daily_filter_worker, 
            self.cb_daily_filter_shift
        ]
        for widget in filter_widgets:
            widget.bind("<Return>", lambda e: self.update_daily_usage_view())
            widget.bind("<<ComboboxSelected>>", lambda e: self.update_daily_usage_view())

        for date_widget in [self.ent_daily_start_date, self.ent_daily_end_date]:
            date_widget.bind("<<DateEntrySelected>>", lambda e: self.update_daily_usage_view())
            try:
                date_widget.bind("<Return>", lambda e: self.update_daily_usage_view())
                for child in date_widget.winfo_children():
                    if isinstance(child, (tk.Entry, ttk.Entry)):
                        child.bind("<Return>", lambda e: self.update_daily_usage_view())
            except: pass
        
        # Treeview for daily usage records
        tree_container = ttk.Frame(display_frame)
        tree_container.pack(expand=True, fill='both', padx=5, pady=5)
        
        list_frame = ttk.Frame(tree_container)
        list_frame.pack(fill='both', expand=True)

        # Scrollbars
        vsb = ttk.Scrollbar(list_frame, orient="vertical")
        hsb = ttk.Scrollbar(list_frame, orient="horizontal")
        
        # Treeview with RTK categories and NDT materials
        # Note: Workers 1-10 columns are kept in the 'columns' tuple for data storage,
        # but we will only show a consolidated '작업자' in 'displaycolumns'.
        # Added '(Full작업자)' for Excel export backup.
        columns = ('날짜', '업체명', '적용코드', '현장', '구분', '검사품명', '성적서번호', '작업자', '작업시간', 'OT1', 'OT2', 'OT3', 'OT4', 'OT5', 'OT6', 'OT7', 'OT8', 'OT9', 'OT10', '장비명', '검사방법', '회사코드', '수량', '단위', '단가', '출장비', '일식', '검사비', 'OT시간', 'OT금액', '품목명', '센터미스', '농도', '마킹미스', '필름마크', '취급부주의', '고객불만', '기타', 'RTK총계', '형광자분', '흑색자분', '백색페인트', '침투제', '세척제', '현상제', '형광침투제', '비고', '입력시간', '차량번호', '주행거리', '차량점검', '차량비고', '(Full작업자)')
        self.daily_usage_tree = ttk.Treeview(list_frame, columns=columns, show='headings',
                                              yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Define display columns - hidden (Full작업자) and individual OT columns by default
        # [NEW] Also hide 'OT시간' and '필름매수' by default per user request
        visible_defaults = ['날짜', '업체명', '적용코드', '현장', '구분', '검사품명', '성적서번호', '작업자', '작업시간', '장비명', '검사방법', '수량', '단위', '단가', '출장비', '일식', '검사비', 'OT금액', '품목명', 'RTK총계', '비고']
        self.daily_usage_tree['displaycolumns'] = visible_defaults
        
        # [NEW NDT COLUMNS] Add dynamically
        new_cols = ['작업형태', '조건1', '조건2', '보정계수', '제경비', '기술료', '환산물량', '재료비', '인건비', '검사구분', '조인트수', '불량수', '관경(Inch)']
        c_list = list(columns)
        for nc in new_cols:
            if nc not in c_list:
                c_list.append(nc)
        columns = tuple(c_list)
        
        v_list = list(visible_defaults)
        for nc in new_cols:
            if nc not in v_list:
                v_list.append(nc)
        
        self.daily_usage_tree.configure(columns=columns)
        self.daily_usage_tree['displaycolumns'] = v_list
        
        for nc in new_cols:
            self.daily_usage_tree.heading(nc, text=nc, command=lambda c=nc: self.treeview_sort_column(self.daily_usage_tree, c, False))
            self.daily_usage_tree.column(nc, width=100, minwidth=20, stretch=False, anchor='center')

        
        vsb.config(command=self.daily_usage_tree.yview)
        hsb.config(command=self.daily_usage_tree.xview)
        
        # Column configuration
        col_widths = {
            '날짜': 160, '현장': 130, '작업자': 170, '작업시간': 120,
            'OT1': 140, 'OT2': 140, 'OT3': 140, 'OT4': 140, 'OT5': 140, 'OT6': 140,
            'OT7': 140, 'OT8': 140, 'OT9': 140, 'OT10': 140,
            '차량번호': 120, '주행거리': 100, '차량점검': 200, '차량비고': 150,
            '장비명': 130, '검사방법': 90, '회사코드': 80, '적용코드': 100, '검사품명': 130, '성적서번호': 120,
            '수량': 70, '단위': 60, '단가': 90, '출장비': 90, '업체명': 120, '일식': 80, 
            '검사비': 100, 'OT시간': 80, 'OT금액': 100, '품목명': 210, '센터미스': 70, '농도': 70, '마킹미스': 70, 
            '필름마크': 70, '취급부주의': 70, '고객불만': 70, '기타': 70, 'RTK총계': 80, 
            '형광자분': 80, '흑색자분': 80, '백색페인트': 80, '침투제': 80, '세척제': 80, 
            '현상제': 80, '형광침투제': 80, '비고': 230, '입력시간': 300
        }
        
        for col in columns:
            self.daily_usage_tree.heading(col, text=col, command=lambda c=col: self.treeview_sort_column(self.daily_usage_tree, c, False))
            width = col_widths.get(col, 100)
            self.daily_usage_tree.column(col, width=width, minwidth=20, stretch=False, anchor='center')
        
        # Grid layout for list_frame
        self.daily_usage_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)
        
        # Auto-save column widths when user resizes columns
        # [NEW] Bind worker details popup AND Quick Edit dialog to double-click
        self.daily_usage_tree.bind("<Button-1>", lambda e: self.show_worker_popup(e, self.daily_usage_tree), add="+")
        self.daily_usage_tree.bind("<Double-1>", lambda e: self.show_worker_popup(e, self.daily_usage_tree), add="+")
        self.daily_usage_tree.bind("<Double-1>", lambda e: self.open_edit_daily_usage_dialog(), add="+")
        
        # [NEW] Enable column reordering via drag & drop
        self.enable_tree_column_drag(self.daily_usage_tree)
        
        def save_column_widths(event=None):
            self.save_tab_config()
            
        self.daily_usage_tree.bind('<ButtonRelease-1>', save_column_widths)

        # [NEW] Note Detail Area for long descriptions
        detail_frame = ttk.LabelFrame(display_frame, text="상세 비고 (선택된 항목)")
        detail_frame.pack(fill='x', padx=5, pady=(0, 5))
        
        self.txt_daily_note_detail = tk.Text(detail_frame, height=3, font=('Malgun Gothic', 10), wrap='word', bg='#F9F9F9')
        self.txt_daily_note_detail.pack(fill='both', expand=True, padx=5, pady=5)
        self.txt_daily_note_detail.config(state='disabled')

        # Selection binding to update detail area
        self.daily_usage_tree.bind('<<TreeviewSelect>>', self._on_daily_usage_select)
        
        # Optimize treeview scroll region to prevent unnecessary scrollbars
        def optimize_treeview_scroll_region():
            try:
                if hasattr(self, 'daily_usage_tree'):
                    self.daily_usage_tree.update_idletasks()
                    
                    # Get treeview content dimensions
                    total_items = len(self.daily_usage_tree.get_children())
                    if total_items > 0:
                        # Calculate approximate content height
                        item_height = 25  # Approximate height per row
                        content_height = total_items * item_height + 50  # Add some padding
                        
                        # Get treeview dimensions
                        tree_width = self.daily_usage_tree.winfo_width()
                        tree_height = self.daily_usage_tree.winfo_height()
                        
                        # Set scroll region to content size if smaller than treeview
                        if content_height < tree_height:
                            self.daily_usage_tree.configure(yscrollcommand=(0, 0, tree_width, content_height))
                            print(f"Treeview scroll region optimized: (0, 0, {tree_width}, {content_height})")
                        else:
                            # Use full treeview dimensions
                            self.daily_usage_tree.configure(yscrollcommand=(0, 0, tree_width, tree_height))
                            print(f"Treeview scroll region set to full size: (0, 0, {tree_width}, {tree_height})")
            except Exception as e:
                print(f"Error optimizing treeview scroll region: {e}")
        
        # Apply optimization after a short delay
        self.root.after(500, optimize_treeview_scroll_region)


    def _on_daily_usage_select(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import _on_daily_usage_select_impl
        return _on_daily_usage_select_impl(self, *args, **kwargs)


    

    def export_budget_estimation(self):
        """Export the estimated budget details to Excel"""
        site = self.cb_budget_site.get()
        if not site:
            messagebox.showwarning("현장 미선택", "사정원가를 내보낼 현장을 선택하거나 입력해주세요.")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"사정예산서_{site}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not file_path:
            return
            
        # Collect Summary Data
        summary_data = {
            'site': site,
            'period': getattr(self, 'ent_budget_period').get() if hasattr(self, 'ent_budget_period') else "",
            'revenue': getattr(self, 'ent_budget_revenue').get() if hasattr(self, 'ent_budget_revenue') else "",
            'unitprice': getattr(self, 'ent_budget_unit_price').get() if hasattr(self, 'ent_budget_unit_price') else "",
            'labor': getattr(self, 'ent_budget_labor').get() if hasattr(self, 'ent_budget_labor') else "",
            'material': getattr(self, 'ent_budget_material').get() if hasattr(self, 'ent_budget_material') else "",
            'expense': getattr(self, 'ent_budget_expense').get() if hasattr(self, 'ent_budget_expense') else "",
            'outsource': getattr(self, 'ent_budget_outsource').get() if hasattr(self, 'ent_budget_outsource') else "",
            'profit': getattr(self, 'ent_budget_profit').get() if hasattr(self, 'ent_budget_profit') else "",
            'margin': getattr(self, 'ent_budget_margin').get() if hasattr(self, 'ent_budget_margin') else ""
        }
        
        # Details
        labor_data = self.labor_detail_widget.get_data() if hasattr(self, 'labor_detail_widget') else {}
        
        # Material Details - need to augment with name/spec/unit from default_items
        raw_mat = self.material_detail_widget.get_data() if hasattr(self, 'material_detail_widget') else []
        material_data = []
        defaults = self.material_detail_widget.default_items if hasattr(self, 'material_detail_widget') else []
        for i, row in enumerate(raw_mat):
            if i < len(defaults):
                name, spec, unit, _ = defaults[i]
                row['name'] = name
                row['spec'] = spec
                row['unit'] = unit
                material_data.append(row)
                
        expense_data = self.expense_detail_widget.get_data() if hasattr(self, 'expense_detail_widget') else {}
        
        try:
            from site_apps.central.src.utils.export_helper import export_budget_estimation_to_excel
            if export_budget_estimation_to_excel(file_path, summary_data, labor_data, material_data, expense_data):
                messagebox.showinfo("내보내기 완료", f"사정예산서가 성공적으로 저장되었습니다.\n{file_path}")
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("오류", f"엑셀 저장 중 오류가 발생했습니다.\n{e}")

    def export_budget_sales_status(self):
        """Export the currently filtered budget sales status to Excel"""
        if not hasattr(self, 'budget_view_tree') or not self.budget_view_tree.get_children():
            messagebox.showwarning("데이터 없음", "내보낼 데이터가 없습니다. 먼저 조회를 해주세요.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"매출현황_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not file_path:
            return

        # Collect data from treeview
        data = []
        columns = [self.budget_view_tree.heading(col)['text'] for col in self.budget_view_tree['columns']]
        for item in self.budget_view_tree.get_children():
            data.append(self.budget_view_tree.item(item)['values'])
            
        df = pd.DataFrame(data, columns=columns)
        
        # Add summary row at the bottom
        summary_row = {col: '' for col in columns}
        summary_row['날짜'] = '합계'
        if hasattr(self, 'lbl_bv_actual_income'):
            summary_row['검사단가'] = self.lbl_bv_actual_income.cget('text').replace(' 원', '').replace(',', '').strip()
        
        # Compute OT/Total directly from columns since Budget KPIs might show differently
        ot_idx = columns.index('OT합계') if 'OT합계' in columns else -1
        total_idx = columns.index('합계') if '합계' in columns else -1
        
        sum_ot = 0
        sum_tot = 0
        for _d in data:
            if ot_idx >= 0:
                try: sum_ot += float(str(_d[ot_idx]).replace(',', ''))
                except: pass
            if total_idx >= 0:
                try: sum_tot += float(str(_d[total_idx]).replace(',', ''))
                except: pass

        summary_row['OT합계'] = f"{sum_ot:g}"
        summary_row['합계'] = f"{sum_tot:g}"
        
        # Add the summary row to the data list for the DataFrame
        data.append([summary_row.get(col, '') for col in columns])
        
        df = pd.DataFrame(data, columns=columns)
        
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='SalesStatus', index=False)
                # Auto-adjust columns
                worksheet = writer.sheets['SalesStatus']
                for i, col in enumerate(df.columns):
                    max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                    worksheet.column_dimensions[chr(65 + i)].width = max_len
            
            messagebox.showinfo("내보내기 완료", "매출현황이 엑셀로 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 저장 중 오류가 발생했습니다: {e}")

    def setup_ndt_billing_tab(self):
        # [NEW] Popout Control Frame
        ctrl_frame = ttk.Frame(self.tab_ndt_billing)
        ctrl_frame.pack(fill='x', padx=5, pady=5)
        ttk.Label(ctrl_frame, text="💡 이 탭을 별도의 창으로 분리하여 듀얼 모니터에서 작업할 수 있습니다.", foreground='#00529B').pack(side='left')
        ttk.Button(ctrl_frame, text="🔍 팝업창으로 열기", command=self.open_detached_ndt_billing_view).pack(side='right')
        
        # Embed the NDT Calculator as a frame inside this tab
        self.ndt_calculator = NDTCalculatorTab(self.tab_ndt_billing, main_app=self)
        self.ndt_calculator.pack(fill='both', expand=True)

    def open_detached_ndt_billing_view(self):
        """기성 정산(NDT) 탭을 별도의 팝업창으로 엽니다."""
        if hasattr(self, 'detached_windows') and 'ndt_billing' in self.detached_windows and self.detached_windows['ndt_billing']['window'].winfo_exists():
            self.detached_windows['ndt_billing']['window'].lift()
            return
            
        if not hasattr(self, 'detached_windows'):
            self.detached_windows = {}
            
        popup = tk.Toplevel(self.root)
        popup.title("💰 기성 정산 (NDT) (팝업)")
        popup.geometry("1600x900")
        
        self.detached_windows['ndt_billing'] = {'window': popup}
        
        popup_calc = NDTCalculatorTab(popup, main_app=self)
        popup_calc.pack(fill='both', expand=True, padx=5, pady=5)

    def setup_budget_tab(self):
        """Setup the project execution budget management tab with detailed labor cost breakdown"""
        # KPI Summary Panel [NEW]
        kpi_frame = tk.Frame(self.tab_budget, background="#ffffff", highlightthickness=1, highlightbackground="#e5e7eb")
        kpi_frame.pack(fill='x', padx=10, pady=(5, 0))
        
        kpis = [
            ("계약금액", "lbl_kpi_rev", "#374151"),
            ("실행원가", "lbl_kpi_cost", "#374151"),
            ("영업이익", "lbl_kpi_profit", "#10b981"),
            ("이익률", "lbl_kpi_margin", "#10b981")
        ]
        
        for i, (label, attr, color) in enumerate(kpis):
            f = tk.Frame(kpi_frame, background="#ffffff")
            f.pack(side='left', expand=True, fill='both', pady=10)
            if i > 0:
                tk.Frame(kpi_frame, width=1, background="#e5e7eb").pack(side='left', fill='y', pady=10)
            
            tk.Label(f, text=label, font=("Malgun Gothic", 9), background="#ffffff", foreground="#6b7280").pack()
            lbl_val = tk.Label(f, text="0원", font=("Malgun Gothic", 12, "bold"), background="#ffffff", foreground=color)
            lbl_val.pack()
            setattr(self, attr, lbl_val)

        # Fixed Button Frame [NEW]
        top_btn_frame = ttk.Frame(self.tab_budget)
        top_btn_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Button(top_btn_frame, text="실적 데이터로 채우기", command=self.fill_budget_from_actuals).pack(side='left', padx=5)
        ttk.Button(top_btn_frame, text="저장/수정", command=self.save_budget_entry).pack(side='left', padx=5)
        ttk.Button(top_btn_frame, text="삭제", command=self.delete_budget_entry).pack(side='left', padx=5)
        ttk.Button(top_btn_frame, text="🧹 초기화", command=self.clear_budget_form).pack(side='left', padx=5)
        
        # Pop-out button
        ttk.Button(top_btn_frame, text="🔍 팝업창으로 열기", command=self.open_detached_budget_view).pack(side='left', padx=15)
        
        ttk.Button(top_btn_frame, text="사정원가 출력", command=self.export_budget_estimation).pack(side='right', padx=10)
        ttk.Button(top_btn_frame, text="엑셀 내보내기", command=self.export_budget_sales_status).pack(side='right', padx=10)

        main_paned = ttk.PanedWindow(self.tab_budget, orient='vertical')
        main_paned.pack(fill='both', expand=True, padx=10, pady=(0, 5))
        
        # --- 상단: Scrollable Form area ---
        top_container = ttk.Frame(main_paned)
        main_paned.add(top_container, weight=2)
        
        canvas = tk.Canvas(top_container, highlightthickness=0)
        v_scroll = ttk.Scrollbar(top_container, orient="vertical", command=canvas.yview)
        form_scrollable = ttk.Frame(canvas)
        
        form_scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        _canvas_win = canvas.create_window((0, 0), window=form_scrollable, anchor="nw")
        canvas.configure(yscrollcommand=v_scroll.set)
        
        # [FIX] 우측 빈 공간 제거: Canvas 너비 변경 시 내부 프레임 너비를 동기화
        def _on_canvas_resize(event):
            canvas.itemconfigure(_canvas_win, width=event.width)
        canvas.bind("<Configure>", _on_canvas_resize)
        
        # Scroll binding is handled globally in MaterialManager.__init__
        
        canvas.pack(side="left", fill="both", expand=True)
        v_scroll.pack(side="right", fill="y")
        
        # 1. Main Budget Form
        form_frame = ttk.LabelFrame(form_scrollable, text="실행예산 입력/수정", padding=10)
        form_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        # Grid layout for form fields
        # Grid layout headers
        headers = ["항목", "사전예산(계획)", "사후원가(실적)", "잔여예산(차액)"]
        for col, h in enumerate(headers):
            ttk.Label(form_frame, text=h, font=('Malgun Gothic', 9, 'bold')).grid(row=0, column=col, padx=10, pady=5)
            
        rows = [
            ("현장명", "cb_budget_site", None, None),
            ("공사기간 (일)", "ent_budget_period", "ent_budget_actual_period", "ent_budget_diff_period"),
            ("계약금액(Revenue) (원)", "ent_budget_revenue", "ent_budget_actual_revenue", "ent_budget_diff_revenue"),
            ("매출금액(UnitPrice) (원)", "ent_budget_unit_price", "ent_budget_actual_unit_price", "ent_budget_diff_unit_price"),
            ("실행 노무비(Labor) (원)", "ent_budget_labor", "ent_budget_actual_labor", "ent_budget_diff_labor"),
            ("실행 재료비(Material) (원)", "ent_budget_material", "ent_budget_actual_material", "ent_budget_diff_material"),
            ("실행 경비(Expense) (원)", "ent_budget_expense", "ent_budget_actual_expense", "ent_budget_diff_expense"),
            ("실행 외주비(Outsource) (원)", "ent_budget_outsource", "ent_budget_actual_outsource", "ent_budget_diff_outsource"),
            ("영업이익(Profit) (원)", "ent_budget_profit", "ent_budget_actual_profit", "ent_budget_diff_profit"),
            ("이익률(%)", "ent_budget_margin", "ent_budget_actual_margin", "ent_budget_diff_margin"),
            ("비고", "ent_budget_note", "ent_budget_actual_note", None)
        ]

        self.budget_widgets = {}
        for r_idx, (label, w_plan, w_actual, w_diff) in enumerate(rows, start=1):
            ttk.Label(form_frame, text=label).grid(row=r_idx, column=0, sticky='e', padx=10, pady=2)
            if w_plan == "cb_budget_site":
                w = ttk.Combobox(form_frame, width=20)
                w['values'] = getattr(self, 'sites', [])
                w.grid(row=r_idx, column=1, sticky='ew', padx=5, pady=2, columnspan=3)
            else:
                w = ttk.Entry(form_frame, width=20)
                w.grid(row=r_idx, column=1, sticky='ew', padx=5, pady=2)
            setattr(self, w_plan, w)
            self.budget_widgets[w_plan] = w
            
            if w_actual:
                w_a = ttk.Entry(form_frame, width=20)
                w_a.grid(row=r_idx, column=2, sticky='ew', padx=5, pady=2)
                setattr(self, w_actual, w_a)
                self.budget_widgets[w_actual] = w_a
                
            if w_diff:
                w_d = ttk.Entry(form_frame, width=20, state='readonly')
                w_d.grid(row=r_idx, column=3, sticky='ew', padx=5, pady=2)
                setattr(self, w_diff, w_d)
                self.budget_widgets[w_diff] = w_d

        def _on_budget_site_selected(e):
            # [FIX] 현장명을 가장 먼저 캡처 - focus_set() 호출 전에 해야 값이 유지됨
            selected_site = self.cb_budget_site.get().strip()
            # 폼 로드 (현장명 먼저 캡처했으므로 안전)
            self._load_budget_to_form(selected_site, silent=True)
            # 포커스 이동은 폼 로드 완료 후에 (after로 지연)
            self.root.after(50, self.ent_budget_revenue.focus_set)
        self.cb_budget_site.bind('<<ComboboxSelected>>', _on_budget_site_selected)

        # 1.5. 상세 탭 컨테이너 분리 (사전예산 / 사후원가)
        self.detail_notebook = ttk.Notebook(form_scrollable)
        self.detail_notebook.pack(fill='both', expand=True, pady=(0, 10), padx=5)
        
        self.tab_planned = ttk.Frame(self.detail_notebook)
        self.tab_actual = ttk.Frame(self.detail_notebook)
        
        self.detail_notebook.add(self.tab_planned, text='사전예산 상세 (Planned)')
        self.detail_notebook.add(self.tab_actual, text='사후원가(실적) 상세 (Actuals)')

        # ===== [ 사전예산 (Planned) 상세 위젯 ] =====
        labor_detail_frame = ttk.LabelFrame(self.tab_planned, text="인건비 상세 (Labor Cost Detail)", padding=10)
        labor_detail_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        def on_labor_change(total):
            self.ent_budget_labor.delete(0, tk.END)
            self.ent_budget_labor.insert(0, f"{total:,.0f}")
            self._update_budget_kpis()
            
        self.labor_detail_widget = LaborCostDetailWidget(labor_detail_frame, on_change_callback=on_labor_change)
        self.labor_detail_widget.pack(fill='x', expand=True)

        material_detail_frame = ttk.LabelFrame(self.tab_planned, text="재료비 상세 (Material Cost Detail)", padding=10)
        material_detail_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        def on_material_change(total):
            self.ent_budget_material.delete(0, tk.END)
            self.ent_budget_material.insert(0, f"{total:,.0f}")
            self._update_budget_kpis()
            
        self.material_detail_widget = MaterialCostDetailWidget(material_detail_frame, on_change_callback=on_material_change)
        self.material_detail_widget.pack(fill='x', expand=True)

        expense_detail_frame = ttk.LabelFrame(self.tab_planned, text="경비 및 이익 상세 (Expense & Profit Detail)", padding=10)
        expense_detail_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        def on_expense_change(exp_total, outsource_total, op_profit):
            self.ent_budget_expense.delete(0, tk.END)
            self.ent_budget_expense.insert(0, f"{exp_total:,.0f}")
            self.ent_budget_outsource.delete(0, tk.END)
            self.ent_budget_outsource.insert(0, f"{outsource_total:,.0f}")
            if hasattr(self, 'ent_budget_profit'):
                rev = get_rev()
                margin = (op_profit / rev * 100) if rev > 0 else 0.0
                self.ent_budget_profit.delete(0, tk.END)
                self.ent_budget_profit.insert(0, f"{op_profit:,.0f} ({margin:.1f}%)")
            self._update_budget_kpis()

        def get_lab():
            try: return float(self.ent_budget_labor.get().replace(',', '') or 0)
            except: return 0.0
        def get_mat():
            try: return float(self.ent_budget_material.get().replace(',', '') or 0)
            except: return 0.0
        def get_rev():
            try: 
                rev = float(self.ent_budget_revenue.get().replace(',', '') or 0)
                if rev == 0 and hasattr(self, 'ent_budget_unit_price'):
                    rev = float(self.ent_budget_unit_price.get().replace(',', '') or 0)
                return rev
            except: return 0.0
            
        self.expense_detail_widget = ExpenseProfitDetailWidget(
            expense_detail_frame, 
            on_change_callback=on_expense_change,
            get_labor_total_func=get_lab,
            get_material_total_func=get_mat,
            get_revenue_func=get_rev
        )
        self.expense_detail_widget.pack(fill='x', expand=True)
        
        # ===== [ 사후원가 (Actual) 상세 위젯 ] =====
        a_labor_detail_frame = ttk.LabelFrame(self.tab_actual, text="실적 인건비 상세 (Actual Labor Detail)", padding=10)
        a_labor_detail_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        def on_actual_labor_change(total):
            self.ent_budget_actual_labor.delete(0, tk.END)
            self.ent_budget_actual_labor.insert(0, f"{total:,.0f}")
            self._update_budget_kpis()
            
        self.actual_labor_detail_widget = LaborCostDetailWidget(a_labor_detail_frame, on_change_callback=on_actual_labor_change)
        self.actual_labor_detail_widget.pack(fill='x', expand=True)

        a_material_detail_frame = ttk.LabelFrame(self.tab_actual, text="실적 재료비 상세 (Actual Material Detail)", padding=10)
        a_material_detail_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        def on_actual_material_change(total):
            self.ent_budget_actual_material.delete(0, tk.END)
            self.ent_budget_actual_material.insert(0, f"{total:,.0f}")
            self._update_budget_kpis()
            
        self.actual_material_detail_widget = MaterialCostDetailWidget(a_material_detail_frame, on_change_callback=on_actual_material_change)
        self.actual_material_detail_widget.pack(fill='x', expand=True)

        a_expense_detail_frame = ttk.LabelFrame(self.tab_actual, text="실적 경비 및 이익 상세 (Actual Expense & Profit)", padding=10)
        a_expense_detail_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        def on_actual_expense_change(exp_total, outsource_total, op_profit):
            self.ent_budget_actual_expense.delete(0, tk.END)
            self.ent_budget_actual_expense.insert(0, f"{exp_total:,.0f}")
            self.ent_budget_actual_outsource.delete(0, tk.END)
            self.ent_budget_actual_outsource.insert(0, f"{outsource_total:,.0f}")
            if hasattr(self, 'ent_budget_actual_profit'):
                rev = get_actual_rev()
                margin = (op_profit / rev * 100) if rev > 0 else 0.0
                self.ent_budget_actual_profit.delete(0, tk.END)
                self.ent_budget_actual_profit.insert(0, f"{op_profit:,.0f} ({margin:.1f}%)")
            self._update_budget_kpis()

        def get_actual_lab():
            try: return float(self.ent_budget_actual_labor.get().replace(',', '') or 0)
            except: return 0.0
        def get_actual_mat():
            try: return float(self.ent_budget_actual_material.get().replace(',', '') or 0)
            except: return 0.0
        def get_actual_rev():
            try: 
                rev = float(self.ent_budget_actual_unit_price.get().replace(',', '') or 0)
                if rev == 0 and hasattr(self, 'ent_budget_actual_revenue'):
                    rev = float(self.ent_budget_actual_revenue.get().replace(',', '') or 0)
                return rev
            except: return 0.0
            
        self.actual_expense_detail_widget = ExpenseProfitDetailWidget(
            a_expense_detail_frame, 
            on_change_callback=on_actual_expense_change,
            get_labor_total_func=get_actual_lab,
            get_material_total_func=get_actual_mat,
            get_revenue_func=get_actual_rev
        )
        self.actual_expense_detail_widget.pack(fill='x', expand=True)
        
        # [NEW] 도급액 실시간 입력에 따른 이익금액 재계산 바인딩
        if hasattr(self, 'ent_budget_revenue'):
            self.ent_budget_revenue.bind('<KeyRelease>', lambda e: self.expense_detail_widget.calculate_all())
            self.ent_budget_revenue.bind('<FocusOut>', lambda e: self.format_entry_with_commas(e, self.ent_budget_revenue), add='+')
            self.ent_budget_revenue.bind('<Return>', lambda e: self.format_entry_with_commas(e, self.ent_budget_revenue), add='+')
            
        if hasattr(self, 'ent_budget_unit_price'):
            self.ent_budget_unit_price.bind('<FocusOut>', lambda e: self.format_entry_with_commas(e, self.ent_budget_unit_price), add='+')
            self.ent_budget_unit_price.bind('<Return>', lambda e: self.format_entry_with_commas(e, self.ent_budget_unit_price), add='+')

        # [MOVED to top_btn_frame]

        # ===============================================================
        # --- 하단: 현장별 일일사용량 조회 프레임 ---
        # ===============================================================
        bottom_container = ttk.Frame(main_paned)
        main_paned.add(bottom_container, weight=3)

        # 필터 영역
        bottom_filter = ttk.Frame(bottom_container)
        bottom_filter.pack(fill='x', pady=(5, 2))

        ttk.Label(bottom_filter, text="현장별 실적 조회", font=('Malgun Gothic', 10, 'bold')).pack(side='left', padx=(0, 15))

        ttk.Label(bottom_filter, text="현장:").pack(side='left', padx=5)
        self.cb_budget_view_site = ttk.Combobox(bottom_filter, width=20, state='readonly')
        site_values = ['전체'] + sorted([str(s).strip() for s in (self.sites or []) if str(s).strip()])
        self.cb_budget_view_site['values'] = site_values
        self.cb_budget_view_site.pack(side='left', padx=5)
        self.cb_budget_view_site.set('전체')
        self.cb_budget_view_site.bind('<<ComboboxSelected>>', lambda e: self.update_budget_site_view())

        ttk.Label(bottom_filter, text="시작일:").pack(side='left', padx=(10, 2))
        self.budget_view_start = DateEntry(bottom_filter, width=12, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        self.budget_view_start.pack(side='left', padx=2)
        today = datetime.datetime.now()
        self.budget_view_start.set_date(datetime.date(2024, 1, 1))

        ttk.Label(bottom_filter, text="종료일:").pack(side='left', padx=(5, 2))
        self.budget_view_end = DateEntry(bottom_filter, width=12, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        self.budget_view_end.pack(side='left', padx=2)
        self.budget_view_end.set_date(today)

        ttk.Button(bottom_filter, text="조회", command=self.update_budget_site_view).pack(side='left', padx=8)
        ttk.Button(bottom_filter, text="예산 수정/불러오기",
                   command=lambda: self._load_budget_to_form(self.cb_budget_view_site.get())).pack(side='left', padx=4)
        ttk.Button(bottom_filter, text="컬럼 추가", command=self.add_budget_view_custom_column).pack(side='left', padx=4)
        ttk.Button(bottom_filter, text="컬럼 관리", command=self.show_budget_view_column_dialog).pack(side='left', padx=4)
        ttk.Button(bottom_filter, text="엑셀 내보내기", command=self.export_budget_sales_status).pack(side='right', padx=10)

        # KPI 요약 행 제거 (예산입력 5번 영업이익률로 대체)
        # Treeview (일일사용량 코럼)
        tree_outer = ttk.Frame(bottom_container)
        tree_outer.pack(fill='both', expand=True)

        # 전체 컬럼: 기본 표시 컬럼 + 사용자가 "추가" 할 수 있는 숨김 기본 컬럼 + 사용자정의 컬럼
        self.budget_view_builtin_cols = (
            'Date', 'Site', '장비명', '검사방법', '작업자', '품목명',
            '총기성액', '청구인건비', '청구재료비', '제경비', '기술료',
            '총지출액', '자재원가', 'OT합계', '출장비', '일식', 
            '예상이윤', '목표이익률', '현재이익률', '검사량', '단가', '검사단가', '합계', 
            '침투제', '세척제', '현상제', '자재단가', '비고',
            '입력시간', '차량번호', '주행거리', '차량점검', '차량비고', 'MaterialID'
        )
        self.budget_view_builtin_width_map = {
            'Date': 90, 'Site': 120, '장비명': 100, '검사방법': 80, '검사량': 70, '단가': 80,
            '검사단가': 90, '출장비': 80, '일식': 70, 'OT합계': 80, '침투제': 70, '세척제': 70, '현상제': 70, '자재단가': 90,
            '합계': 100, '비고': 150, '작업자': 170, '품목명': 180, '입력시간': 140,
            '차량번호': 110, '주행거리': 100, '차량점검': 120, '차량비고': 140, 'MaterialID': 90,
            '총기성액': 100, '청구인건비': 90, '청구재료비': 90, '제경비': 90, '기술료': 90,
            '총지출액': 100, '자재원가': 90, '예상이윤': 100, '목표이익률': 80, '현재이익률': 80
        }
        self.budget_view_builtin_head_map = {
            'Date': '날짜', 'Site': '현장', '장비명': '장비', '검사방법': '검사방법', '검사량': '수량', '단가': '단가',
            '검사단가': '매출금액', '출장비': '출장비', '일식': '일식', 'OT합계': 'OT지급액', '침투제': '침투제', '세척제': '세척제', '현상제': '현상제',
            '자재단가': '자재단가', '합계': '합계', '비고': '비고', '작업자': '작업자', '품목명': '품목명',
            '입력시간': '입력시간', '차량번호': '차량번호', '주행거리': '주행거리', '차량점검': '차량점검',
            '차량비고': '차량비고', 'MaterialID': '자재ID',
            '총기성액': '총기성액', '청구인건비': '인건비(청구)', '청구재료비': '재료비(청구)', '제경비': '제경비', '기술료': '기술료',
            '총지출액': '총지출액', '자재원가': '자재원가', '예상이윤': '예상이윤', '목표이익률': '목표이익률(%)', '현재이익률': '현재이익률(%)'
        }
        self.budget_view_heading_aliases = getattr(self, 'budget_view_heading_aliases', {})
        self.budget_view_custom_columns = getattr(self, 'budget_view_custom_columns', [])
        self.budget_view_default_cols = (
            'Date', 'Site', '장비명', '검사방법', '작업자', '품목명',
            '총기성액', '청구인건비', '청구재료비', '제경비', '기술료',
            '총지출액', '자재원가', 'OT합계', '출장비', '일식', 
            '예상이윤', '목표이익률', '현재이익률', '비고'
        )

        vsb = ttk.Scrollbar(tree_outer, orient='vertical')
        hsb = ttk.Scrollbar(tree_outer, orient='horizontal')
        self.budget_view_tree = ttk.Treeview(tree_outer, columns=self.budget_view_builtin_cols, show='headings',
                                              yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._refresh_budget_view_tree_columns(reload_data=False)

        vsb.config(command=self.budget_view_tree.yview)
        hsb.config(command=self.budget_view_tree.xview)
        self.budget_view_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        tree_outer.grid_rowconfigure(0, weight=1)
        tree_outer.grid_columnconfigure(0, weight=1)
        self.budget_view_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())
        self.enable_tree_column_drag(self.budget_view_tree, context_menu_handler=self._show_budget_view_heading_context_menu)

        # ── ESC: 현장별 탭 내 모든 입력 위젯에서 포커스 해제 ──────────────
        def _esc_to_root(e=None):
            self.root.focus_set()
            return "break"

        def _bind_esc_recursive(widget):
            try:
                cls = widget.winfo_class()
                # Toplevel(달력 팝업 등) 내부는 절대 순회하지 않음
                # → 순회만으로도 winfo_children() 가 Toplevel을 화면에 표시시킬 수 있음
                if cls == 'Toplevel':
                    return
                if cls in ('TEntry', 'Entry', 'TCombobox'):
                    widget.bind('<Escape>', _esc_to_root, add='+')
            except Exception:
                pass
            for child in widget.winfo_children():
                _bind_esc_recursive(child)

        # after_idle: 모든 위젯 생성 완료 후 바인딩
        self.root.after_idle(lambda: _bind_esc_recursive(self.tab_daily_usage))
        
        # 초기 데이터 로드 (첫 현장 자동 선택 후 표시)
        self.root.after(100, lambda: self.update_budget_site_view())

    def update_budget_view(self):
        """Refresh the budget inquiry view to reflect updated budget data."""
        if hasattr(self, 'cb_budget_view_site'):
            # Only refresh if a site is actually being viewed
            if self.cb_budget_view_site.get():
                self.update_budget_site_view()

    def _refresh_budget_view_tree_columns(self, reload_data=False):
        """Rebuild budget site performance tree column structure."""
        if not hasattr(self, 'budget_view_tree'):
            return

        custom_cols = getattr(self, 'budget_view_custom_columns', []) or []
        # [FIX] built-in 컬럼과 이름/키가 중복되는 사용자 컬럼 정리
        builtin_keys = set(getattr(self, 'budget_view_builtin_cols', []) or [])
        builtin_headings = set((getattr(self, 'budget_view_builtin_head_map', {}) or {}).values())
        sanitized_custom_cols = []
        seen_custom_keys = set()
        seen_custom_names = set()
        for c in custom_cols:
            key = str(c.get('key', '')).strip()
            name = str(c.get('name', key)).strip()
            if not key:
                continue
            # built-in 과 키 충돌/이름 충돌이면 제거
            if key in builtin_keys or name in builtin_headings:
                continue
            # custom 내부 중복도 제거
            if key in seen_custom_keys or name in seen_custom_names:
                continue
            seen_custom_keys.add(key)
            seen_custom_names.add(name)
            sanitized_custom_cols.append(c)
        if len(sanitized_custom_cols) != len(custom_cols):
            self.budget_view_custom_columns = sanitized_custom_cols
            custom_cols = sanitized_custom_cols

        custom_keys = [c.get('key') for c in custom_cols if c.get('key')]
        self.budget_view_cols = tuple(list(self.budget_view_builtin_cols) + custom_keys)

        head_map = dict(getattr(self, 'budget_view_builtin_head_map', {}))
        width_map = dict(getattr(self, 'budget_view_builtin_width_map', {}))
        for col in custom_cols:
            key = col.get('key')
            if not key:
                continue
            head_map[key] = col.get('name', key)
            width_map[key] = int(col.get('width', 120) or 120)

        alias_map = getattr(self, 'budget_view_heading_aliases', {}) or {}
        self.budget_view_tree.configure(columns=self.budget_view_cols)
        for col in self.budget_view_cols:
            heading_text = alias_map.get(col, head_map.get(col, col))
            self.budget_view_tree.heading(col, text=heading_text,
                                          command=lambda c=col: self.treeview_sort_column(self.budget_view_tree, c, False))
            self.budget_view_tree.column(col, width=width_map.get(col, 100), minwidth=40, anchor='center', stretch=False)

        visible_source = getattr(self, 'budget_view_visible_cols', []) or list(self.budget_view_default_cols)
        visible = [c for c in visible_source if c in self.budget_view_cols]
        self.budget_view_tree['displaycolumns'] = visible if visible else self.budget_view_default_cols
        self.budget_view_visible_cols = list(self.budget_view_tree['displaycolumns'])

        if reload_data:
            self.update_budget_site_view()

    def add_budget_view_custom_column(self):
        """Add a custom column to budget site performance view."""
        name = simpledialog.askstring("컬럼 추가", "새 컬럼 이름을 입력하세요:", parent=self.root)
        if not name or not str(name).strip():
            return
        name = str(name).strip()

        # [FIX] built-in/기존 사용자 컬럼과 이름 중복 방지
        builtin_names = set((getattr(self, 'budget_view_builtin_head_map', {}) or {}).values())
        existing_custom_names = {str(c.get('name', '')).strip() for c in (getattr(self, 'budget_view_custom_columns', []) or [])}
        if name in builtin_names or name in existing_custom_names:
            messagebox.showwarning("중복 컬럼", f"'{name}' 컬럼은 이미 존재합니다.")
            return

        default_value = simpledialog.askstring("컬럼 내용", f"'{name}' 컬럼의 기본 내용을 입력하세요(선택):", parent=self.root)

        existing = set(getattr(self, 'budget_view_cols', []))
        key_base = re.sub(r'[^0-9A-Za-z가-힣_]+', '_', name).strip('_') or '사용자컬럼'
        key = key_base
        idx = 1
        while key in existing:
            idx += 1
            key = f"{key_base}_{idx}"

        custom_cols = getattr(self, 'budget_view_custom_columns', []) or []
        custom_cols.append({'key': key, 'name': name, 'default': default_value or '', 'width': 120})
        self.budget_view_custom_columns = custom_cols

        visible = list(getattr(self, 'budget_view_visible_cols', []) or list(self.budget_view_default_cols))
        if key not in visible:
            visible.append(key)
        self.budget_view_visible_cols = visible

        self._refresh_budget_view_tree_columns(reload_data=True)
        self.save_tab_config()

    def rename_budget_view_column(self, col_key):
        """Rename a budget site performance column heading."""
        if not col_key:
            return
        current_name = self.budget_view_tree.heading(col_key)['text'] if hasattr(self, 'budget_view_tree') else col_key
        new_name = simpledialog.askstring("컬럼 이름 변경", "새 컬럼 이름을 입력하세요:", initialvalue=current_name, parent=self.root)
        if not new_name or not str(new_name).strip():
            return
        new_name = str(new_name).strip()

        # Custom column: rename the stored name itself
        updated = False
        for col in getattr(self, 'budget_view_custom_columns', []) or []:
            if col.get('key') == col_key:
                col['name'] = new_name
                updated = True
                break
        if not updated:
            aliases = getattr(self, 'budget_view_heading_aliases', {}) or {}
            default_name = getattr(self, 'budget_view_builtin_head_map', {}).get(col_key, col_key)
            if new_name == default_name:
                aliases.pop(col_key, None)
            else:
                aliases[col_key] = new_name
            self.budget_view_heading_aliases = aliases

        self._refresh_budget_view_tree_columns(reload_data=False)
        self.save_tab_config()

    def set_budget_view_custom_column_content(self, col_key):
        """Change default content for a custom column."""
        custom_cols = getattr(self, 'budget_view_custom_columns', []) or []
        target = next((c for c in custom_cols if c.get('key') == col_key), None)
        if not target:
            return
        new_value = simpledialog.askstring("컬럼 내용 설정", f"'{target.get('name', col_key)}' 컬럼의 기본 내용을 입력하세요:",
                                           initialvalue=target.get('default', ''), parent=self.root)
        if new_value is None:
            return
        target['default'] = new_value
        self._refresh_budget_view_tree_columns(reload_data=True)
        self.save_tab_config()

    def delete_budget_view_custom_column(self, col_key):
        """Delete a custom budget column."""
        custom_cols = getattr(self, 'budget_view_custom_columns', []) or []
        target = next((c for c in custom_cols if c.get('key') == col_key), None)
        if not target:
            return
        if not messagebox.askyesno("컬럼 삭제", f"'{target.get('name', col_key)}' 컬럼을 삭제하시겠습니까?", parent=self.root):
            return

        self.budget_view_custom_columns = [c for c in custom_cols if c.get('key') != col_key]
        self.budget_view_visible_cols = [c for c in getattr(self, 'budget_view_visible_cols', []) if c != col_key]
        aliases = getattr(self, 'budget_view_heading_aliases', {}) or {}
        aliases.pop(col_key, None)
        self.budget_view_heading_aliases = aliases
        self._refresh_budget_view_tree_columns(reload_data=True)
        self.save_tab_config()

    def show_budget_view_column_dialog(self):
        """Open dialog to show/hide columns in budget site performance view"""
        if not hasattr(self, 'budget_view_tree'):
            return

        all_cols = list(self.budget_view_tree['columns'])
        active_cols = self.budget_view_tree['displaycolumns']
        if not active_cols or active_cols == ('#all'):
            active_cols = all_cols

        dialog_cols = [(col, self.budget_view_tree.heading(col)['text']) for col in all_cols]
        dialog = ColumnSelectionDialog(self.root, dialog_cols, title="현장별 실적 표시 컬럼 관리")
        for col, var in dialog.vars.items():
            var.set(col in active_cols)

        dialog.wait_window()

        if dialog.result is not None:
            # 필수 컬럼 보정: 날짜/현장은 항상 유지
            final_selection = list(dialog.result)
            for mandatory in ('Date', 'Site'):
                if mandatory in all_cols and mandatory not in final_selection:
                    final_selection.append(mandatory)

            sorted_selection = [c for c in all_cols if c in final_selection]
            self.budget_view_tree['displaycolumns'] = sorted_selection if sorted_selection else all_cols
            self.budget_view_visible_cols = list(self.budget_view_tree['displaycolumns'])
            self.save_tab_config()

    def _show_budget_view_heading_context_menu(self, event):
        """공사탭 현장별 실적 Treeview 헤더 우클릭 메뉴"""
        if not hasattr(self, 'budget_view_tree'):
            return
        tree = self.budget_view_tree
        try:
            if tree.identify_region(event.x, event.y) != 'heading':
                return
        except Exception:
            return

        column_id = tree.identify_column(event.x)
        col_key = self._get_column_name_from_id(tree, column_id)
        if not col_key:
            return
        col_text = tree.heading(col_key)['text']
        custom_keys = {c.get('key') for c in (getattr(self, 'budget_view_custom_columns', []) or [])}

        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label=f"⬅️ '{col_text}' 왼쪽으로 이동", command=lambda: self._move_column_visual(tree, column_id, -1))
        menu.add_command(label=f"➡️ '{col_text}' 오른쪽으로 이동", command=lambda: self._move_column_visual(tree, column_id, 1))
        menu.add_separator()
        menu.add_command(label=f"✏️ '{col_text}' 이름 변경...", command=lambda: self.rename_budget_view_column(col_key))
        menu.add_command(label="➕ 사용자 컬럼 추가...", command=self.add_budget_view_custom_column)
        if col_key in custom_keys:
            menu.add_command(label=f"📝 '{col_text}' 내용 설정...", command=lambda: self.set_budget_view_custom_column_content(col_key))
            menu.add_command(label=f"🗑️ '{col_text}' 사용자 컬럼 삭제", command=lambda: self.delete_budget_view_custom_column(col_key))
        menu.add_separator()
        menu.add_command(label="⚙️ 컬럼 관리(추가/삭제)...", command=self.show_budget_view_column_dialog)
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def update_budget_site_view(self):
        """현장별 일일사용량 데이터를 하단 Treeview에 표시하고 KPI를 갱신한다."""
        try:
            if not hasattr(self, 'budget_view_tree'):
                return

            for item in self.budget_view_tree.get_children():
                self.budget_view_tree.delete(item)

            if not hasattr(self, 'cb_budget_view_site'):
                return
            site = self.cb_budget_view_site.get().strip()
            
            # [NEW] Track logically unique entries to avoid double-counting materials from split-row records
            processed_entry_ids = set()

            # [FIX] Do NOT return early if df is empty, we still need to run column hiding logic
            if self.daily_usage_df.empty:
                df_empty = True
            else:
                df_empty = False

            # --- 날짜 필터 ---
            start_date = self.budget_view_start.get_date()
            end_date   = self.budget_view_end.get_date()
            df = self.daily_usage_df.copy()
            # 구버전 컬럼명 호환
            if 'Site' not in df.columns and '현장' in df.columns:
                df['Site'] = df['현장']
            if 'Date' not in df.columns and '날짜' in df.columns:
                df['Date'] = df['날짜']
            if 'Site' not in df.columns:
                df_empty = True
            
            if not df_empty:
                df['_site_norm'] = df['Site'].astype(str).str.strip()

            # 날짜 파싱(실패 데이터는 제외 대신 전체 누락 방지를 위해 후단에서 완화)
            if 'Date' in df.columns:
                # Use normalized Timestamps for robust comparison with tk date objects
                df['_date'] = pd.to_datetime(df['Date'], errors='coerce').dt.normalize()
                start_ts = pd.to_datetime(start_date).normalize()
                end_ts = pd.to_datetime(end_date).normalize()
                date_mask = (df['_date'] >= start_ts) & (df['_date'] <= end_ts)
                # [STRICT] Exclude any rows with invalid/missing dates (NaT)
                date_mask = date_mask & df['_date'].notna()
            else:
                date_mask = pd.Series([True] * len(df), index=df.index)

            if site and site != '전체':
                site_mask = (df['_site_norm'] == site)
            else:
                site_mask = pd.Series([True] * len(df), index=df.index)

            mask = date_mask & site_mask
            df = df[mask].copy()

            # --- 자재 원가 맵 ---
            mat_id_cost_map = {}
            if 'MaterialID' in self.materials_df.columns and '원가' in self.materials_df.columns:
                mat_id_cost_map = self.materials_df.set_index('MaterialID')['원가'].fillna(0).to_dict()

            def _f(val):
                if pd.isna(val) or str(val).lower() == 'nan': return 0.0
                try: 
                    s = str(val).replace(',', '').strip()
                    if not s: return 0.0
                    num = float(s)
                    return 0.0 if np.isnan(num) else num
                except: return 0.0

            total_income = 0.0   # 검사비 + OT 합계
            total_expense = 0.0  # 출장비 + 일식
            total_net_revenue = 0.0 # 검사비 합계
            total_sum     = 0.0
            total_mat_cost = 0.0 # 자재 실적 사용료 합계
            total_ndt_penetrant = 0.0 # NDT 침투제 합계
            total_ndt_cleaner = 0.0 # NDT 세척제 합계
            total_ndt_developer = 0.0 # NDT 현상제 합계
            
            # [NEW] Enhanced Budget Totals
            sys_total_revenue = 0.0
            sys_total_labor = 0.0
            sys_total_mat_bill = 0.0
            sys_total_overhead = 0.0
            sys_total_tech = 0.0
            sys_total_expense_cost = 0.0
            sys_total_profit = 0.0

            import re as _re
            
            # [NEW] Robust is_active helper for smart column hiding
            def is_active(val):
                if val is None: return False
                s = str(val).strip().lower()
                if s in ('', '0', '0.0', '0.00', 'nan', 'none', '-', '0원', '0.0원', '0시간'):
                    return False
                try:
                    # Remove common units for numeric check
                    v_raw = s.replace(',', '').replace('원', '').replace('시간', '').strip()
                    if not v_raw: return False
                    v = float(v_raw)
                    return abs(v) > 0.001
                except:
                    return bool(s)

            # [NEW] Smart Column Hiding tracking
            mandatory_cols = {'Date', 'Site', '합계', '비고'}
            has_data_map = {col: False for col in self.budget_view_cols if col not in mandatory_cols}

            def _clean_str(val):
                if pd.isna(val):
                    return ''
                s = str(val).strip()
                return '' if s.lower() in ('nan', 'none') else s

            def _first_text(_row, keys, default=''):
                """여러 후보 컬럼 중 첫 유효 문자열 반환 (구버전 컬럼명 호환)."""
                for k in keys:
                    if k in _row.index:
                        v = _clean_str(_row.get(k, ''))
                        if v:
                            return v
                return default

            def _ndt_val(_row, *keys):
                for k in keys:
                    if k in _row.index:
                        return _f(_row.get(k, 0))
                return 0.0

            for _, row in df.iterrows():
                # [NEW] Logical entry deduplication for the display list (Treeview)
                e_date = str(row.get('Date', '')).split(' ')[0]
                e_site = str(row.get('Site', '')).strip()
                raw_time = str(row.get('EntryTime', '')).strip()
                e_time = raw_time[:16] if len(raw_time) > 16 else raw_time
                e_uid = f"{e_date}|{e_site}|{e_time}"
                
                if e_uid in processed_entry_ids:
                    continue # Skip redundant rows from being displayed in the list
                processed_entry_ids.add(e_uid)

                # [FIX] OT 컬럼에서 금액 추출 로직 강화 (NaN 방지)
                parsed_ots = []
                for i in range(1, 11):
                    raw_ot = row.get('OT' if i == 1 else f'OT{i}', 0)
                    if pd.isna(raw_ot) or not str(raw_ot).strip() or str(raw_ot).lower() == 'nan':
                        parsed_ots.append(0.0)
                        continue
                    
                    v_str = str(raw_ot).strip()
                    if '(' in v_str and '원)' in v_str:
                        try:
                            amt_str = v_str.split('(')[1].split('원')[0].replace(',', '').strip()
                            parsed_ots.append(_f(amt_str))
                        except: parsed_ots.append(0.0)
                    else:
                        parsed_ots.append(_f(v_str))
                
                ot_sum = sum(parsed_ots)

                net    = _f(row.get('검사비', 0))
                travel = _f(row.get('출장비', 0))
                meal   = _f(row.get('일식', 0))
                qty    = _f(row.get('검사량', 0))
                price  = _f(row.get('단가', 0))

                # 자재원가
                test_method = str(row.get('검사방법', '')).upper()
                if 'RT' in test_method:
                    usage = _f(row.get('Usage', 0))
                else:
                    usage = _f(row.get('Usage', 0))

                # [NEW] Track logically unique entries to avoid double-counting materials from split-row records
                # Based on Date + Site + EntryTime (best UID) or Date + Site + Note (as fallback)
                e_time = str(row.get('EntryTime', '')).strip()
                e_site = str(row.get('Site', row.get('현장', ''))).strip()
                e_date = str(row.get('Date', row.get('날짜', ''))).strip()
                e_note = str(row.get('Note', row.get('비고', ''))).strip()
                e_uid = f"{e_date}|{e_site}|{e_time}|{e_note}"
                
                is_new_logical_entry = (e_uid not in processed_entry_ids)
                processed_entry_ids.add(e_uid)

                # 일일 탭 NDT 입력값(침투제/세척제/현상제)
                # [FIX] ONLY sum consumables for the first row of each logical entry to prevent inflation
                ndt_penetrant = 0.0
                ndt_cleaner = 0.0
                ndt_developer = 0.0
                
                if is_new_logical_entry:
                    ndt_penetrant = _ndt_val(row, 'NDT_침투제', '침투제', 'NDT_침투액', '침투액')
                    ndt_cleaner = _ndt_val(row, 'NDT_세척제', '세척제', 'NDT_세척액', '세척액')
                    ndt_developer = _ndt_val(row, 'NDT_현상제', '현상제', 'NDT_현상액', '현상액')

                    total_ndt_penetrant += ndt_penetrant
                    total_ndt_cleaner += ndt_cleaner
                    total_ndt_developer += ndt_developer

                unit_cost = mat_id_cost_map.get(row.get('MaterialID'), 0)
                mat_cost = usage * float(unit_cost)
                total_mat_cost += mat_cost

                # [NEW] Calculate Revenue and Profit per row
                nd_labor = _f(row.get('인건비', 0))
                nd_mat = _f(row.get('재료비', 0))
                nd_overhead = _f(row.get('제경비', 0))
                nd_tech = _f(row.get('기술료', 0))
                
                # If NDT fields exist, use them. Otherwise fallback to standard net revenue.
                row_revenue = nd_labor + nd_mat + nd_overhead + nd_tech
                if row_revenue <= 0:
                    row_revenue = net

                row_expense_cost = travel + meal + ot_sum + mat_cost
                row_profit = row_revenue - row_expense_cost
                row_margin = (row_profit / row_revenue * 100) if row_revenue > 0 else 0.0

                # --- Financial Totals (Sum for ALL rows) ---
                total_net_revenue += net
                total_expense += (travel + meal)
                total_income += (net + ot_sum)
                total_sum += (net + travel + meal)
                
                # [NEW] Accumulate system totals
                sys_total_revenue += row_revenue
                sys_total_labor += nd_labor
                sys_total_mat_bill += nd_mat
                sys_total_overhead += nd_overhead
                sys_total_tech += nd_tech
                sys_total_expense_cost += row_expense_cost
                sys_total_profit += row_profit

                # 작업자 통합
                worker_cols = ['User'] + [f'User{i}' for i in range(2, 11)]
                raw_workers = []
                for wcol in worker_cols:
                    wv = _clean_str(row.get(wcol, ''))
                    if wv and wv not in raw_workers:
                        raw_workers.append(wv)
                consolidated_workers = ', '.join(raw_workers)

                # 품목명 표시명
                try:
                    mat_name = self.get_material_display_name(row.get('MaterialID')) if pd.notna(row.get('MaterialID')) else ''
                except Exception:
                    mat_name = ''

                row_map = {
                    'Date': _clean_str(row.get('Date', '')),
                    'Site': _clean_str(row.get('Site', '')),
                    '장비명': _clean_str(row.get('장비명', '')),
                    '검사방법': _clean_str(row.get('검사방법', '')),
                    '검사량': f"{qty:g}",
                    '단가': f"{price:,.0f}",
                    '검사단가': f"{net:,.0f}",
                    '출장비': f"{travel:,.0f}",
                    '일식': f"{meal:,.0f}",
                    'OT합계': f"{ot_sum:,.0f}",
                    '침투제': f"{_f(row.get('NDT_침투제', row.get('침투제', 0))):g}",
                    '세척제': f"{_f(row.get('NDT_세척제', row.get('세척제', 0))):g}",
                    '현상제': f"{_f(row.get('NDT_현상제', row.get('현상제', 0))):g}",
                    '자재단가': f"{unit_cost:,.0f}",
                    '합계': f"{(net + travel + meal):,.0f}",
                    '비고': _clean_str(row.get('비고', row.get('Note', ''))),
                    '작업자': consolidated_workers,
                    '품목명': mat_name,
                    '입력시간': _clean_str(row.get('EntryTime', '')),
                    '차량번호': _first_text(row, ['차량번호', 'vehicle_info', 'VehicleInfo']),
                    '주행거리': _first_text(row, ['주행거리', 'mileage', 'Mileage']),
                    '차량점검': _first_text(row, ['차량점검', '차량 점검', 'vehicle_checks', 'inspection', 'Inspection']),
                    '차량비고': _first_text(row, ['차량비고', '차량 비고', 'remarks', 'vehicle_remarks', 'Remark']),
                    'MaterialID': _clean_str(row.get('MaterialID', '')),
                    'Usage': _clean_str(row.get('Usage', '')),
                    '총기성액': f"{row_revenue:,.0f}" if row_revenue else '',
                    '청구인건비': f"{nd_labor:,.0f}" if nd_labor else '',
                    '청구재료비': f"{nd_mat:,.0f}" if nd_mat else '',
                    '제경비': f"{nd_overhead:,.0f}" if nd_overhead else '',
                    '기술료': f"{nd_tech:,.0f}" if nd_tech else '',
                    '총지출액': f"{row_expense_cost:,.0f}" if row_expense_cost else '',
                    '자재원가': f"{mat_cost:,.0f}" if mat_cost else '',
                    '예상이윤': f"{row_profit:,.0f}" if row_profit else '',
                    '목표이익률': '',
                    '현재이익률': f"{row_margin:.1f}%" if row_revenue > 0 else '',
                }
                for custom_col in getattr(self, 'budget_view_custom_columns', []) or []:
                    key = custom_col.get('key')
                    if key:
                        row_map[key] = custom_col.get('default', '')
                
                # [NEW] Mark columns that have non-zero/non-empty data using robust is_active
                for col, val in row_map.items():
                    if col in has_data_map and not has_data_map[col]:
                        if is_active(val):
                            has_data_map[col] = True

                self.budget_view_tree.insert('', tk.END, values=tuple(row_map.get(col, '') for col in self.budget_view_cols))

                try:
                    equip_name = str(row.get('장비명', ''))
                    m_name_local = str(row.get('검사방법', ''))
                    date_key = str(row.get('Date', ''))
                    if 'RT' in m_name_local or 'RT' in equip_name or 'RT' in mat_name:
                        rt_dates.add(date_key)
                    
                    # [NEW] Generic equipment tracking by name
                    if equip_name and equip_name.lower() != 'nan':
                        if equip_name not in equip_dates_map:
                            equip_dates_map[equip_name] = set()
                        equip_dates_map[equip_name].add(date_key)
                except Exception:
                    pass

            # --- 합계 행 추가 ---
            if df.shape[0] > 0:
                total_row_map = {
                    'Date': f"[합계]",
                    'Site': f"{site}",
                    '장비명': '',
                    '검사방법': '',
                    '검사량': '',
                    '단가': '',
                    '검사단가': f"{total_net_revenue:,.0f}",
                    '출장비': f"{total_expense:,.0f}",
                    '일식': '',
                    'OT합계': '',
                    '자재사용량': '',
                    '침투제': f"{total_ndt_penetrant:g}",
                    '세척제': f"{total_ndt_cleaner:g}",
                    '현상제': f"{total_ndt_developer:g}",
                    '자재단가': '',
                    '합계': f"{total_sum:,.0f}",
                    '비고': f"소계: 수입 {total_income:,.0f} 경비 {total_expense:,.0f} 자재원가 {total_mat_cost:,.0f}",
                    '작업자': '',
                    '품목명': '',
                    '입력시간': '',
                    '차량번호': '',
                    '주행거리': '',
                    '차량점검': '',
                    '차량비고': '',
                    'MaterialID': '',
                    'Usage': '',
                    '총기성액': f"{sys_total_revenue:,.0f}" if sys_total_revenue else '',
                    '청구인건비': f"{sys_total_labor:,.0f}" if sys_total_labor else '',
                    '청구재료비': f"{sys_total_mat_bill:,.0f}" if sys_total_mat_bill else '',
                    '제경비': f"{sys_total_overhead:,.0f}" if sys_total_overhead else '',
                    '기술료': f"{sys_total_tech:,.0f}" if sys_total_tech else '',
                    '총지출액': f"{sys_total_expense_cost:,.0f}" if sys_total_expense_cost else '',
                    '자재원가': f"{total_mat_cost:,.0f}" if total_mat_cost else '',
                    '예상이윤': f"{sys_total_profit:,.0f}" if sys_total_profit else '',
                    '목표이익률': '',
                    '현재이익률': f"{(sys_total_profit / sys_total_revenue * 100):.1f}%" if sys_total_revenue > 0 else '',
                }
                for custom_col in getattr(self, 'budget_view_custom_columns', []) or []:
                    key = custom_col.get('key')
                    if key:
                        total_row_map[key] = ''
                self.budget_view_tree.insert('', tk.END, values=tuple(total_row_map.get(col, '') for col in self.budget_view_cols), tags=('total',))

            # --- [NEW] 예산 대비 실적 요약 행 추가 ---
            if df.shape[0] > 0 and site and site != '전체' and hasattr(self, 'budget_df') and not self.budget_df.empty:
                budget_match = self.budget_df[self.budget_df['Site'] == site]
                if not budget_match.empty:
                    b_row = budget_match.iloc[0]
                    
                    # Fetch Budget Values
                    def to_f_b(key):
                        try: 
                            v = str(b_row.get(key, 0)).replace(',', '').strip()
                            return float(v) if v else 0.0
                        except: return 0.0
                    
                    b_rev = to_f_b('Revenue')
                    b_labor = to_f_b('LaborCost')
                    b_mat = to_f_b('MaterialCost')
                    b_exp = to_f_b('Expense')
                    b_profit = to_f_b('Profit')
                    
                    # Progress / Variance Calculations
                    rev_progress = (total_income / b_rev * 100) if b_rev > 0 else 0
                    mat_progress = (total_mat_cost / b_mat * 100) if b_mat > 0 else 0
                    exp_progress = (total_expense / b_exp * 100) if b_exp > 0 else 0
                    
                    # Final Net Profit Achievement
                    actual_profit = total_income - total_expense - total_mat_cost
                    b_margin = (b_profit / b_rev * 100) if b_rev > 0 else 0
                    actual_margin = (sys_total_profit / sys_total_revenue * 100) if sys_total_revenue > 0 else 0
                    
                    comparison_row_map = {
                        'Date': f"[예산대비]",
                        'Site': f"달성률: {rev_progress:.1f}%",
                        '장비명': '',
                        '검사방법': '',
                        '검사량': '',
                        '단가': '',
                        '검사단가': f"예산 {b_rev:,.0f}",
                        '출장비': f"경비 {exp_progress:.1f}%",
                        '일식': '',
                        'OT합계': '',
                        '자재사용량': f"자재 {mat_progress:.1f}%",
                        '침투제': '',
                        '세척제': '',
                        '현상제': '',
                        '자재단가': '',
                        '합계': f"{actual_profit:,.0f}",
                        '비고': f"목표이익: {b_profit:,.0f} | 현재이익: {actual_profit:,.0f} ({actual_profit - b_profit:+,.0f})",
                        '목표이익률': f"{b_margin:.1f}%",
                        '현재이익률': f"{actual_margin:.1f}%",
                        '작업자': '',
                        '품목명': '',
                        '입력시간': '',
                        '차량번호': '',
                        '주행거리': '',
                        '차량점검': '',
                        '차량비고': '',
                        'MaterialID': '',
                        'Usage': '',
                        '총기성액': '', '청구인건비': '', '청구재료비': '', '제경비': '', '기술료': '',
                        '총지출액': '', '자재원가': '', '예상이윤': ''
                    }
                    for custom_col in getattr(self, 'budget_view_custom_columns', []) or []:
                        key = custom_col.get('key')
                        if key:
                            comparison_row_map[key] = ''
                    
                    # Insert with a distinct tag for styling
                    self.budget_view_tree.tag_configure('comparison', background='#E3F2FD', font=('Malgun Gothic', 10, 'bold'))
                    self.budget_view_tree.insert('', tk.END, values=tuple(comparison_row_map.get(col, '') for col in self.budget_view_cols), tags=('comparison',))

            # --- [NEW] Smart Column Hiding Application ---
            active_cols = [col for col in self.budget_view_cols 
                           if col in mandatory_cols or has_data_map.get(col, False)]
            self.budget_view_tree['displaycolumns'] = active_cols
            
            # --- Treeview 태그 스타일 (합계 행 강조) ---

            if not self.budget_view_tree.tag_configure('total'):
                self.budget_view_tree.tag_configure('total', background='#ffff00', foreground='#000000')
        except Exception as e:
            print(f"ERROR in update_budget_site_view: {e}")


    def _update_budget_kpis(self):
        """Update the top KPI summary labels based on current budget form values"""
        try:
            def _get_val(widget_name):
                try: 
                    w = getattr(self, widget_name, None)
                    if w:
                        v = str(w.get()).replace(',', '').split(' ')[0]
                        return float(v or 0)
                    return 0.0
                except: return 0.0
                
            def _set_val(widget_name, val, is_percent=False):
                w = getattr(self, widget_name, None)
                if w:
                    st = w.cget('state')
                    if st == 'readonly': w.config(state='normal')
                    w.delete(0, 'end')
                    w.insert(0, f"{val:,.1f}" if is_percent else f"{val:,.0f}")
                    if st == 'readonly': w.config(state='readonly')

            def _calc_profit(prefix):
                rev = _get_val(f'{prefix}revenue')
                if rev == 0: rev = _get_val(f'{prefix}unit_price')
                lab = _get_val(f'{prefix}labor')
                mat = _get_val(f'{prefix}material')
                exp = _get_val(f'{prefix}expense')
                out = _get_val(f'{prefix}outsource')
                
                if prefix == 'ent_budget_' and hasattr(self, 'expense_detail_widget') and hasattr(self.expense_detail_widget, 'lbl_grand_total_cost'):
                    try:
                        raw_t = self.expense_detail_widget.lbl_grand_total_cost.cget('text')
                        tc = float("".join(c for c in raw_t if c.isdigit() or c == '.') or 0)
                    except: tc = lab + mat + exp + out
                else:
                    tc = lab + mat + exp + out
                    
                prof = rev - tc
                mar = (prof / rev * 100) if rev > 0 else 0.0
                _set_val(f'{prefix}profit', prof)
                _set_val(f'{prefix}margin', mar, True)
                return rev, lab, mat, exp, out, prof, mar
                
            p_r, p_l, p_m, p_e, p_o, p_p, p_mg = _calc_profit('ent_budget_')
            a_r, a_l, a_m, a_e, a_o, a_p, a_mg = _calc_profit('ent_budget_actual_')
            
            _set_val('ent_budget_diff_revenue', p_r - a_r)
            _set_val('ent_budget_diff_unit_price', _get_val('ent_budget_unit_price') - _get_val('ent_budget_actual_unit_price'))
            _set_val('ent_budget_diff_labor', p_l - a_l)
            _set_val('ent_budget_diff_material', p_m - a_m)
            _set_val('ent_budget_diff_expense', p_e - a_e)
            _set_val('ent_budget_diff_outsource', p_o - a_o)
            _set_val('ent_budget_diff_profit', p_p - a_p)
            _set_val('ent_budget_diff_margin', p_mg - a_mg, True)
            
            if hasattr(self, 'lbl_kpi_rev'):
                self.lbl_kpi_rev.config(text=f"{p_r:,.0f}원")
                self.lbl_kpi_cost.config(text=f"{p_l+p_m+p_e+p_o:,.0f}원")
                self.lbl_kpi_profit.config(text=f"{p_p:,.0f}원", foreground="#ef4444" if p_p < 0 else "#10b981")
                self.lbl_kpi_margin.config(text=f"{p_mg:.1f}%", foreground="#ef4444" if p_mg < 0 else "#10b981")
            self.root.update_idletasks()
        except Exception as e:
            print(f"DEBUG: Budget KPI sync error: {e}")

    def _load_budget_to_form(self, site, silent=False):
        """선택한 현장의 예산서를 상단 입력 폼에 로드한다."""
        # [NEW] Clear current form before loading new site data to prevent residual data
        self.clear_budget_form()

        # [FIX] clear_budget_form()이 cb_budget_site를 지우므로 즉시 복원
        # budget_df에 해당 현장이 없더라도 콤보박스에 선택한 현장명은 유지해야 함
        if site and hasattr(self, 'cb_budget_site'):
            self.cb_budget_site.set(site)

        if not site:
            if not silent: messagebox.showwarning("현장 미선택", "현장을 먼저 선택해주세요.")
            return
        if not hasattr(self, 'budget_df') or self.budget_df.empty:
            if not silent: messagebox.showinfo("데이터 없음", f"'{site}' 현장의 저장된 예산서가 없습니다.")
            return
        match = self.budget_df[self.budget_df['Site'] == site]
        if match.empty:
            if not silent: messagebox.showinfo("데이터 없음", f"'{site}' 현장의 저장된 예산서가 없습니다.")
            return

        row = match.iloc[0]
        self.cb_budget_site.set(site)
        
        mappings = [
            ('ent_budget_period', 'Period'), ('ent_budget_actual_period', 'Actual_Period'),
            ('ent_budget_revenue', 'Revenue'), ('ent_budget_unit_price','UnitPrice'),
            ('ent_budget_labor', 'LaborCost'), ('ent_budget_material', 'MaterialCost'),
            ('ent_budget_expense', 'Expense'), ('ent_budget_outsource', 'OutsourceCost'),
            ('ent_budget_profit', 'Profit'), ('ent_budget_note', 'Note'),
            ('ent_budget_actual_revenue', 'Actual_Revenue'), ('ent_budget_actual_unit_price','Actual_UnitPrice'),
            ('ent_budget_actual_labor', 'Actual_LaborCost'), ('ent_budget_actual_material', 'Actual_MaterialCost'),
            ('ent_budget_actual_expense', 'Actual_Expense'), ('ent_budget_actual_outsource', 'Actual_OutsourceCost'),
            ('ent_budget_actual_profit', 'Actual_Profit'), ('ent_budget_actual_note', 'Actual_Note')
        ]
        
        for attr, col in mappings:
            w = getattr(self, attr, None)
            if w and col in match.columns:
                st = w.cget('state')
                if st == 'readonly': w.config(state='normal')
                w.delete(0, 'end')
                val = row[col]
                if not pd.isna(val) and str(val).lower() != 'nan':
                    if 'note' not in attr.lower():
                        try:
                            # Try to convert to float for formatting
                            f_val = float(str(val).replace(',', ''))
                            w.insert(0, f"{f_val:,.0f}")
                        except ValueError:
                            w.insert(0, str(val))
                    else:
                        w.insert(0, str(val))
                if st == 'readonly': w.config(state='readonly')

        import json as _json
        for json_col, widget_attr in [('LaborDetail',   'labor_detail_widget'),
                                       ('MaterialDetail','material_detail_widget'),
                                       ('ExpenseDetail', 'expense_detail_widget'),
                                       ('Actual_LaborDetail',   'actual_labor_detail_widget'),
                                       ('Actual_MaterialDetail','actual_material_detail_widget'),
                                       ('Actual_ExpenseDetail', 'actual_expense_detail_widget')]:
            jdata = row.get(json_col, '')
            widget = getattr(self, widget_attr, None)
            if widget:
                # [FIX] 사후원가(실적) 위젯인데 DB에 저장된 내역이 없는 경우, 사전예산 내역의 텍스트 구조를 복사 (인원/수량은 비움)
                if json_col.startswith('Actual_') and (not jdata or str(jdata).strip() in ['', '{}', '[]']):
                    planned_col = json_col.replace('Actual_', '')
                    p_jdata = row.get(planned_col, '')
                    if p_jdata and str(p_jdata).strip() not in ['', '{}', '[]']:
                        try:
                            p_data = _json.loads(p_jdata)
                            # 인원 및 수량 초기화
                            if 'Labor' in json_col and isinstance(p_data, dict):
                                for k, v in p_data.items():
                                    if isinstance(v, dict):
                                        v['personnel'] = ''
                                        v['period'] = ''
                            elif 'Material' in json_col and isinstance(p_data, list):
                                for v in p_data:
                                    if isinstance(v, dict):
                                        v['qty'] = ''
                            elif 'Expense' in json_col and isinstance(p_data, dict):
                                for v_list in p_data.values():
                                    if isinstance(v_list, list):
                                        for v in v_list:
                                            if isinstance(v, dict):
                                                if 'qty' in v: v['qty'] = ''
                                                if 'days' in v: v['days'] = ''
                            widget.set_data(p_data)
                            continue
                        except Exception as e: 
                            print(f"DEBUG: Failed to copy planned structure to actual widget: {e}")
                
                # 정상 로드 로직
                if jdata and str(jdata).strip() not in ['', '{}', '[]']:
                    try:
                        widget.set_data(_json.loads(jdata))
                    except:
                        widget.reset()
                else:
                    widget.reset()

        if not silent: messagebox.showinfo("로드 완료", f"'{site}' 현장의 예산서를 불러왔습니다.")
        # [STABILITY] Update KPI summary after load
        self._update_budget_kpis()

    def fill_budget_from_actuals(self):
        """하단 필터(현장, 기간) 기준으로 실제 실적 데이터를 집계하여 상단 예산 폼에 자동 입력한다."""
        site = self.cb_budget_view_site.get().strip()
        if not site:
            messagebox.showwarning("현장 미선택", "하단 '현장별 실적 조회' 영역에서 현장을 먼저 선택해주세요.")
            return

        start_ts = pd.Timestamp(self.budget_view_start.get_date())
        end_ts = pd.Timestamp(self.budget_view_end.get_date())

        if self.daily_usage_df.empty:
            messagebox.showinfo("데이터 없음", "집계할 실적 데이터가 없습니다.")
            return

        # 1. 필터링
        df = self.daily_usage_df.copy()
        df['_date_ts'] = pd.to_datetime(df['Date'], errors='coerce')
        if site == '전체':
            mask = (df['_date_ts'] >= start_ts) & (df['_date_ts'] <= end_ts)
        else:
            mask = (df['_date_ts'] >= start_ts) & (df['_date_ts'] <= end_ts) & (df['Site'] == site)
        df = df[mask].copy()

        if df.empty:
            s_date = self.budget_view_start.get_date()
            e_date = self.budget_view_end.get_date()
            messagebox.showinfo("데이터 없음", f"선택한 기간 ({s_date} ~ {e_date}) 동안 '{site}' 현장의 실적이 없습니다.")
            return

        # 2. 집계 초기화
        def _f(val):
            if pd.isna(val) or str(val).lower() == 'nan': return 0.0
            try: 
                s = str(val).replace(',', '').strip()
                if not s: return 0.0
                num = float(s)
                return 0.0 if np.isnan(num) else num
            except: return 0.0

        total_net_revenue = 0.0 # 검사비
        total_travel = 0.0
        total_meal = 0.0
        lab_total = 0.0
        outsource_total = 0.0  # 작업자 없는 행 → 외주비
        
        # --- Labor Aggregation by Rank ---
        ranks = ["이사", "부장", "차장", "과장", "대리", "계장", "주임", "기사"]
        # Track unique dates for each worker to calculate active days correctly
        # Structure: {rank: {worker_name: set(date_strings)}}
        rank_labor_dates = {r: {} for r in ranks}
        
        # --- Special Work (OT) Classification ---
        special_types = ["연장근무", "야간근무", "휴일근무"]
        ot_data = {t: {'names': set(), 'hours': 0} for t in special_types}

        import re as _re
        # Updated pattern to support [Rank], Rank Name, or Name alone with mapping
        # Supports: "[부장] 주진철", "부장 주진철", "주진철"
        rank_pattern = _re.compile(r"\[?(이사|부장|차장|과장|대리|계장|주임|기사)\]?\s*(.*)")
        
        # Default mapping for core workers if rank is missing
        worker_rank_map = {
            "주진철": "부장", "우명광": "대리", "김진환": "주임", "장승대": "계장",
            "김성렬": "주임", "박광복": "부장", "주영광": "과장", "이경재": "주임",
            "황조현": "주임", "김춘호": "차장", "박원준": "과장", "이봉주": "부장",
            "김성현": "대리"
        }

        # 상세 재료비 집계를 위한 맵
        material_usage_sums = [0.0] * 10
        mat_map = {
            '세척': 0, '침투': 1, '현상': 2, '자분': 3, '흑색': 4,
            '필름': 5, '글리세': 6, '현상액': 7, '정착액': 8, '수적': 9
        }

        # 자재 마스터 정보
        mat_id_name_map = {}
        mat_id_cost_map = {}
        if not self.materials_df.empty:
            mat_id_name_map = self.materials_df.set_index('MaterialID')['품목명'].fillna('').to_dict()
            mat_id_cost_map = self.materials_df.set_index('MaterialID')['원가'].fillna(0).to_dict()

        total_mat_cost = 0.0
        rt_inspection_count = 0  # Counter for RT occurrences
        total_film_count = 0.0
        # [NEW] Use unique days instead of row count to avoid inflating days for multiple records
        total_days_count = df['_date_ts'].dt.date.nunique() if not df.empty else 0
        
        # [NEW] Track logically unique entries to avoid double-counting materials from split-row records
        processed_entry_ids = set()
        
        # [NEW] Track unique dates for vehicles and specialized equipment
        starex_dates = set()
        toptruck_dates = set()
        paut_dates = set()
        paut_manual_scanner_dates = set()
        paut_cobra_scanner_dates = set()
        mt_dates = set()
        rt_dates = set()
        # [NEW] Track unique dates for ANY equipment name found in the '장비명' column
        equip_dates_map = {}

        for _, row in df.iterrows():
            # Check weekday
            date_val = _re.sub(r'\s.*', '', str(row.get('Date', '')))
            
            try:
                dt_obj = pd.to_datetime(date_val)
                is_weekday = dt_obj.weekday() < 5 # 0-4: Mon-Fri
            except:
                is_weekday = True

            # [NEW] Enhanced Revenue Logic
            nd_labor = _f(row.get('인건비', 0))
            nd_mat = _f(row.get('재료비', 0))
            nd_overhead = _f(row.get('제경비', 0))
            nd_tech = _f(row.get('기술료', 0))
            has_worker = any(
                str(row.get('User' if i == 1 else f'User{i}', '')).strip() not in ('', 'nan')
                for i in range(1, 11)
            )
            if has_worker:
                lab_total += nd_labor
            else:
                outsource_total += nd_labor
            row_revenue = nd_labor + nd_mat + nd_overhead + nd_tech
            if row_revenue <= 0:
                row_revenue = _f(row.get('검사비', 0))
            total_net_revenue += row_revenue
            total_travel += _f(row.get('출장비', 0))
            total_meal += _f(row.get('일식', 0))

            # [NEW] Identify logically unique entries (shared save event)
            # STRATEGIC FIX: Remove 'Note' from UID because split-rows often have different notes like '(차량 추가 기록)'
            # relying only on Date + Site + EntryTime (minute level) for grouping
            raw_time = str(row.get('EntryTime', '')).strip()
            e_time = raw_time[:16] if len(raw_time) > 16 else raw_time
            e_site = str(row.get('Site', '')).strip()
            
            # UID composite - date_val is already YYYY-MM-DD
            e_uid = f"{date_val}|{e_site}|{e_time}"
            
            is_new_logical_entry = (e_uid not in processed_entry_ids)
            processed_entry_ids.add(e_uid)

            # 명칭 불일치 방지 (세척제/세척액, 침투제/침투액, 현상제/현상액 혼용 대응)
            def get_first_valid(r, keys):
                for k in keys:
                    v = _f(r.get(k, 0))
                    if v > 0: return v
                return 0.0

            # [NOTE] Specialist NDT columns are now handled in the consolidated block below to avoid double counting
            
            # --- Vehicle Tracking ---
            m_id_local = row.get('MaterialID', '')
            # [FIX] If MaterialID is a manual name string (not in master map), use it directly as the name
            m_name_local = str(mat_id_name_map.get(m_id_local, m_id_local)).upper()
            
            # [FIX] Resolve vehicle number from master config to capture '(스타렉스)' tags
            v_no_raw = str(row.get('차량번호', '')).strip()
            vehicles_list = getattr(self, 'vehicles', [])
            vehicle_map = {v.split('(')[0].strip(): v for v in vehicles_list}
            resolved_v_no = vehicle_map.get(v_no_raw, v_no_raw)
            
            # [FIX] Check both Equipment Name (장비명) and Item Name (품목명) for vehicle detection
            car_info = f"{resolved_v_no} {row.get('차량비고', '')} {row.get('장비명', '')} {m_name_local}".strip()
            
            try:
                dt_obj = pd.to_datetime(date_val)
                date_key = dt_obj.date()
            except:
                date_key = date_val
                
            if is_new_logical_entry:
                if '스타렉스' in car_info:
                    starex_dates.add(date_key)
                if '탑차' in car_info:
                    toptruck_dates.add(date_key)

            # PAUT 및 특수장비 작업일수(중복일 제외) 집계
            try:
                method_val = str(row.get('검사방법', '')).upper()
                equip_name = str(row.get('장비명', '')).upper()
                
                # [FIX] Check both Equipment Name and Item Name for PAUT/MT detection
                is_paut = 'PAUT' in method_val or 'PAUT' in equip_name or 'PAUT' in m_name_local
                is_mt = 'MT' in method_val or 'MT' in equip_name or 'MT' in m_name_local or 'YOKE' in equip_name or 'YOKE' in m_name_local
                
                if is_paut:
                    paut_dates.add(date_key)
                    # Specialized Scanner Detection (Robust matching)
                    eq_clean = equip_name.replace(' ', '').upper()
                    m_clean = m_name_local.replace(' ', '').upper()
                    mid_clean = str(m_id_local).replace(' ', '').upper()
                    
                    is_cobra_scanner = ('SCANNER(COBRA)' in eq_clean or 'SCANNER(COBRA)' in m_clean or 'SCANNER(COBRA)' in mid_clean or
                                        ('COBRA' in eq_clean and 'SCANNER' in eq_clean) or 
                                        ('COBRA' in m_clean and 'SCANNER' in m_clean) or
                                        ('COBRA' in mid_clean and 'SCANNER' in mid_clean))

                    # [FIX] If it's a scanner and not COBRA, it's MANUAL
                    is_manual_scanner = ('SCANNER' in eq_clean or 'SCANNER' in m_clean or 'SCANNER' in mid_clean) and not is_cobra_scanner

                    if is_manual_scanner:
                        paut_manual_scanner_dates.add(date_key)
                    if is_cobra_scanner:
                        paut_cobra_scanner_dates.add(date_key)
                        
                if is_mt:
                    mt_dates.add(date_key)
                if 'RT' in method_val or 'RT' in equip_name or 'RT' in m_name_local:
                    rt_dates.add(date_key)
            except Exception:
                pass

            # --- Labor & OT Parsing ---
            shift = str(row.get('Shift', '주간')).strip()
            for i in range(1, 11):
                suffix = '' if i == 1 else str(i)
                raw_worker = str(row.get(f'User{suffix}', '')).strip()
                if not raw_worker or raw_worker.lower() == 'nan':
                    continue
                
                # Rank Aggregation
                match = rank_pattern.search(raw_worker)
                worker_name_only = raw_worker
                rank = None
                
                if match:
                    # Case 1: Rank identified via pattern (e.g., "[부장] 주진철" or "부장 주진철")
                    if match.group(1) in ranks:
                        rank = match.group(1)
                        worker_name_only = match.group(2).strip()
                        if not worker_name_only: # Only rank was present
                             worker_name_only = rank
                
                # Case 2: Fallback to mapping if rank not identified from the string
                if not rank:
                    # Clean potential shift prefixes if they exist (though usually cleaned in records)
                    clean_name = _re.sub(r'\(.*?\)', '', raw_worker).strip()
                    rank = worker_rank_map.get(clean_name)
                    worker_name_only = clean_name
                
                if rank and rank in rank_labor_dates:
                    if worker_name_only not in rank_labor_dates[rank]:
                        rank_labor_dates[rank][worker_name_only] = set()
                    try:
                        # Normalize date key for uniqueness
                        dt_obj = pd.to_datetime(date_val)
                        d_key = dt_obj.date()
                    except:
                        d_key = date_val
                    rank_labor_dates[rank][worker_name_only].add(d_key)
                
                # [FIX] WorkTime 컬럼에서 분류(주간/야간/휴일)와 시간 파싱
                # 형식: "(야간) 4h" or "(주야간) 2.5h" 또는 "8h" (구형식)
                wt_col = f'WorkTime{suffix}'
                wt_val = str(row.get(wt_col, '')).strip()
                
                # WorkTime에서 분류 파싱
                wt_shift = shift  # 기본값: 레코드의 Shift 컬럼
                wt_hours = 0.0
                if wt_val and wt_val.lower() != 'nan':
                    # Parse shift type if present, e.g. "(야간)", "(연장)"
                    shift_match = _re.search(r'\(?(주간|야간|주야간|휴일|연장)\)?', wt_val)
                    if shift_match:
                        wt_shift = shift_match.group(1)
                    
                    # Compute duration calculation
                    marker_pattern = MARKER_PATTERN
                    clean_wt = marker_pattern.sub('', wt_val).replace('(연장)', '').replace('(주야간)', '').replace('(야간)', '').replace('(휴일)', '').replace('(주간)', '').strip()
                    sep_match = _re.search(r'[:\d]\s*([~-])\s*[:\d]', clean_wt)
                    
                    if sep_match:
                        try:
                            sep = sep_match.group(1)
                            st_str, en_str = clean_wt.split(sep)
                            sh, sm = map(int, st_str.split(':'))
                            eh, em = map(int, en_str.split(':'))
                            start_f = sh + sm/60.0
                            end_f = eh + em/60.0
                            if end_f < start_f: end_f += 24
                            wt_hours = end_f - start_f
                        except:
                            pass
                    else:
                        # Fallback: Just a number like "8h" or "8"
                        h_match = _re.search(r'([\d.]+)', clean_wt)
                        if h_match:
                            try: wt_hours = float(h_match.group(1))
                            except: pass
                
                # [FIX] 분류별 OT 상세 집계
                if wt_hours > 0:
                    # 명시적으로 '휴일'이라고 쓰여있거나, 주말인데 기본 교대근무가 아닌 특근성 출근인 경우 (이번엔 단순히 텍스트 명시에 더 의존)
                    if '휴일' in wt_shift:
                        ot_data['휴일근무']['hours'] += wt_hours
                        ot_data['휴일근무']['names'].add(worker_name_only)
                    else:
                        # 평일/주말 공통: 연장이나 야간 명시가 있거나 시간대 기반 계산
                        if sep_match:
                            try:
                                sep = sep_match.group(1)
                                st_str, en_str = clean_wt.split(sep)
                                sh, sm = map(int, st_str.split(':'))
                                eh, em = map(int, en_str.split(':'))
                                start_f = sh + sm/60.0
                                end_f = eh + em/60.0
                                if end_f < start_f: end_f += 24
                                
                                # 현장 기본 퇴근 시간 (17시 기준 혹은 18시 등)
                                # 전체 시간이 주어질 때 명확히 구분 (기본근무: ~18, 연장근무: 18~22, 야간근무: 22~)
                                if end_f > 18:
                                    ot_start = max(start_f, 18.0)
                                    # 18~22: 연장근무
                                    h_over = max(0, min(end_f, 22.0) - ot_start)
                                    # 22~24 : 야간근무
                                    h_night = max(0, min(end_f, 24.0) - max(ot_start, 22.0))
                                    # 24~ : 심야/익일 근무
                                    h_dawn = max(0, end_f - max(ot_start, 24.0))
                                    
                                    if h_over > 0:
                                        ot_data['연장근무']['hours'] += h_over
                                        ot_data['연장근무']['names'].add(worker_name_only)
                                    if h_night > 0:
                                        ot_data['야간근무']['hours'] += h_night
                                        ot_data['야간근무']['names'].add(worker_name_only)
                                    if h_dawn > 0:
                                        try:
                                            # 날짜 확인 (금요일 밤 -> 토요일 새벽인지)
                                            entry_date = pd.to_datetime(row.get('Date'))
                                            if entry_date.weekday() == 4: # 4 = 금요일
                                                ot_data['휴일근무']['hours'] += h_dawn
                                                ot_data['휴일근무']['names'].add(worker_name_only)
                                            else:
                                                ot_data['야간근무']['hours'] += h_dawn
                                                ot_data['야간근무']['names'].add(worker_name_only)
                                        except:
                                            ot_data['야간근무']['hours'] += h_dawn
                                            ot_data['야간근무']['names'].add(worker_name_only)
                            except: pass
                        else:
                            # 시간 범위가 주어지지 않은 숫자(예: 4h, 10h) 처리 로직
                            # 명시적으로 (연장) 태그가 있는 경우
                            if '연장' in wt_shift:
                                ot_data['연장근무']['hours'] += wt_hours
                                ot_data['연장근무']['names'].add(worker_name_only)
                            # 명시적 태그가 없으나 8시간을 초과하는 경우
                            elif wt_hours > 8:
                                h_extra = wt_hours - 8
                                if '야간' in wt_shift or '주야간' in wt_shift:
                                    ot_data['야간근무']['hours'] += h_extra
                                    ot_data['야간근무']['names'].add(worker_name_only)
                                else:
                                    ot_data['연장근무']['hours'] += h_extra
                                    ot_data['연장근무']['names'].add(worker_name_only)


            # --- 자재 사용량 집계 (NDT + 필름) ---
            if is_new_logical_entry:
                m_id = row.get('MaterialID', '')
                m_name = str(mat_id_name_map.get(m_id, '')).strip()
                test_method = str(row.get('검사방법', '')).upper()
                usage = _f(row.get('Usage', 0))
                
                if 'RT' in test_method:
                    rt_inspection_count += 1
                    film_qty = get_first_valid(row, ['필름매수', 'FilmCount', 'Usage', '검사량'])
                    total_film_count += film_qty
                    material_usage_sums[5] += film_qty  # Index 5 = 필름
                    total_mat_cost += film_qty * float(mat_id_cost_map.get(m_id, 0))
                else:
                    # UT/PT/MT 검사: NDT 화학약품 등의 일반 자재 사용
                    # Specialist columns take priority over the generic 'Usage' field
                    ndt_vals = [
                        get_first_valid(row, ['NDT_세척제', '세척제', 'NDT_세척액', '세척액']),
                        get_first_valid(row, ['NDT_침투제', '침투제', 'NDT_침투액', '침투액']),
                        get_first_valid(row, ['NDT_현상제', '현상제', 'NDT_현상액', '현상액']),
                        get_first_valid(row, ['NDT_흑색자분', 'NDT_자분', '흑색자분', 'NDT_페인트']),
                        get_first_valid(row, ['NDT_백색페인트', '백색페인트']) # Placeholder mapping adjustment
                    ]
                    
                    has_specialist_data = any(v > 0 for v in ndt_vals)
                    
                    if has_specialist_data:
                        for idx, v in enumerate(ndt_vals):
                            if idx < 5: material_usage_sums[idx] += v
                        # Cost calculation still needs a MaterialID context, fallback to usage row if available
                        total_mat_cost += usage * float(mat_id_cost_map.get(m_id, 0))
                    elif usage > 0:
                        # Fallback to Usage field matching if specialist columns are empty
                        # 세척제: 청소/세척 용도 → Index 0
                        if '세척제' in m_name or '세척' in m_name or '청소' in m_name:
                            material_usage_sums[0] += usage
                        # 침투제: 침투액 → Index 1
                        elif '침투제' in m_name or '침투액' in m_name:
                            material_usage_sums[1] += usage
                        # 현상제: 현상용 시약 → Index 2
                        elif '현상제' in m_name or '현상액' in m_name:
                            material_usage_sums[2] += usage
                        # 자분: 자기 자분 → Index 3
                        elif '자분' in m_name and '흑색' not in m_name:
                            material_usage_sums[3] += usage
                        # 흑색자분 → Index 4
                        elif '흑색자분' in m_name or ('흑색' in m_name and '자분' in m_name):
                            material_usage_sums[4] += usage
                        # 글리세린 → Index 6
                        elif '글리세' in m_name or '글리세린' in m_name:
                            material_usage_sums[6] += usage
                        # 정착액 → Index 8
                        elif '정착액' in m_name or '정착' in m_name:
                            material_usage_sums[8] += usage
                        # 수적 → Index 9
                        elif '수적' in m_name:
                            material_usage_sums[9] += usage
                        # 기타 PT/MT 화학약품이면 자동 매핑
                        elif any(k in m_name.upper() for k in ['PENETRANT', 'DEVELOPER', 'CLEANER', 'NABAKEM', 'PT', 'MT']):
                            matched = False
                            for keyword, idx in mat_map.items():
                                if keyword.lower() in m_name.lower():
                                    material_usage_sums[idx] += usage
                                    matched = True
                                    break
                            if not matched:
                                material_usage_sums[0] += usage
                        total_mat_cost += usage * float(mat_id_cost_map.get(m_id, 0))

        # 3. 폼에 입력 및 [5. 영업이익] '실행(실적)' 줄에 표시 결과 연동
        # Update Special Unit Prices if LaborCostDetailWidget is available to compute correct revenue
        ot_revenue = 0.0
        if hasattr(self, 'labor_detail_widget'):
            ot_revenue = sum(ot_data[t]['hours'] * _f(self.labor_detail_widget.entries[t]['unit_price'].get()) 
                            for t in special_types if t in self.labor_detail_widget.entries)

        # 3. 폼에 입력 (사후실적 칸만 업데이트)
        self.cb_budget_site.set(site)
        
        # 도급액(Revenue)
        if hasattr(self, 'ent_budget_actual_revenue'):
            self.ent_budget_actual_revenue.delete(0, tk.END)
            self.ent_budget_actual_revenue.insert(0, f"{total_net_revenue:,.0f}")

        # 검사단가
        if hasattr(self, 'ent_budget_actual_unit_price'):
            self.ent_budget_actual_unit_price.delete(0, tk.END)
            self.ent_budget_actual_unit_price.insert(0, f"{total_net_revenue:,.0f}")

        # 재료비
        if hasattr(self, 'ent_budget_actual_material'):
            self.ent_budget_actual_material.delete(0, tk.END)
            self.ent_budget_actual_material.insert(0, f"{np.nan_to_num(total_mat_cost):,.0f}")

        # 경비
        if hasattr(self, 'ent_budget_actual_expense'):
            self.ent_budget_actual_expense.delete(0, tk.END)
            self.ent_budget_actual_expense.insert(0, f"{np.nan_to_num(total_travel + total_meal):,.0f}")

        # 노무비 (lab_total is already calculated from df)
        if hasattr(self, 'ent_budget_actual_labor'):
            self.ent_budget_actual_labor.delete(0, tk.END)
            self.ent_budget_actual_labor.insert(0, f"{np.nan_to_num(lab_total):,.0f}")
            
        # 외주비: 작업자 없는 행의 인건비 합산
        if hasattr(self, 'ent_budget_actual_outsource'):
            self.ent_budget_actual_outsource.delete(0, tk.END)
            self.ent_budget_actual_outsource.insert(0, f"{np.nan_to_num(outsource_total):,.0f}")

        # 4. 실적 인건비 상세 탭 누적 업데이트
        if hasattr(self, 'actual_labor_detail_widget'):
            labor_data = self.actual_labor_detail_widget.get_data()
            planned_labor_data = self.labor_detail_widget.get_data() if hasattr(self, 'labor_detail_widget') else {}
            for rank in ranks:
                if rank in labor_data:
                    u_count = len(rank_labor_dates[rank])
                    t_days = sum(len(dates) for dates in rank_labor_dates[rank].values())
                    if u_count > 0:
                        avg_d = t_days / u_count
                        labor_data[rank]['personnel'] = f"{u_count:g}"
                        labor_data[rank]['period'] = f"{avg_d:g}"
                    else:
                        labor_data[rank]['personnel'] = ""
                        labor_data[rank]['period'] = ""
                    # [FIX] 사전예산 단가 복사
                    if rank in planned_labor_data:
                        labor_data[rank]['unit_price'] = planned_labor_data[rank].get('unit_price', '')
            for stype in special_types:
                if stype in labor_data:
                    u_count = len(ot_data[stype]['names'])
                    t_hours = ot_data[stype]['hours']
                    if u_count > 0:
                        avg_h = t_hours / u_count
                        labor_data[stype]['personnel'] = f"{u_count:g}"
                        labor_data[stype]['period'] = f"{avg_h:g}"
                    else:
                        labor_data[stype]['personnel'] = ""
                        labor_data[stype]['period'] = ""
                    # [FIX] 사전예산 단가 복사
                    if stype in planned_labor_data:
                        labor_data[stype]['unit_price'] = planned_labor_data[stype].get('unit_price', '')
            self.actual_labor_detail_widget.set_data(labor_data)

        # 5. 실적 자재비 상세 탭 누적 업데이트
        if hasattr(self, 'actual_material_detail_widget'):
            current_mat_data = self.actual_material_detail_widget.get_data()
            planned_mat_data = self.material_detail_widget.get_data() if hasattr(self, 'material_detail_widget') else []
            film_chem_qty = (total_film_count / 250.0) if total_film_count > 0 else 0.0
            anti_drop_qty = (total_film_count / 500.0) if total_film_count > 0 else 0.0
            if len(material_usage_sums) > 7:
                material_usage_sums[7] = film_chem_qty
            if len(material_usage_sums) > 8:
                material_usage_sums[8] = film_chem_qty
            if len(material_usage_sums) > 9:
                material_usage_sums[9] = anti_drop_qty

            for i, qty in enumerate(material_usage_sums):
                if i < len(current_mat_data):
                    current_mat_data[i]['qty'] = f"{qty:g}" if qty > 0 else ""
                    # [FIX] 사전예산 단가 복사
                    if i < len(planned_mat_data):
                        current_mat_data[i]['price'] = planned_mat_data[i].get('price', '')
            self.actual_material_detail_widget.set_data(current_mat_data)
        
        # 6. 실적 경비 상세 탭 누적 업데이트
        if hasattr(self, 'actual_expense_detail_widget'):
            current_exp_data = self.actual_expense_detail_widget.get_data()
            planned_exp_data = self.expense_detail_widget.get_data() if hasattr(self, 'expense_detail_widget') else {}
            
            # [NEW] 차량 대수 및 누적 투입 일수 계산 (차량유지비용 및 감가상각비)
            vehicle_dates_map = {}
            vehicles_list = getattr(self, 'vehicles', [])
            vehicle_map = {v.split('(')[0].strip(): v for v in vehicles_list}
            
            for _, row in df.iterrows():
                raw_v_no = str(row.get('차량번호', '')).strip()
                date_val = str(row.get('Date', '')).strip()
                if raw_v_no and raw_v_no.lower() not in ['nan', 'none']:
                    # Split by || or comma in case multiple vehicles are entered in one cell
                    v_list = [v.strip() for v in raw_v_no.replace('||', ',').split(',') if v.strip()]
                    
                    try:
                        dt_obj = pd.to_datetime(date_val)
                        date_key = dt_obj.date()
                    except:
                        date_key = date_val
                        
                    for v_no in v_list:
                        resolved_v = vehicle_map.get(v_no, v_no)
                        if resolved_v not in vehicle_dates_map:
                            vehicle_dates_map[resolved_v] = set()
                        vehicle_dates_map[resolved_v].add(date_key)
            
            total_vehicle_days = sum(len(dates) for dates in vehicle_dates_map.values())
            starex_days = sum(len(dates) for v, dates in vehicle_dates_map.items() if '스타렉스' in v)
            toptruck_days = sum(len(dates) for v, dates in vehicle_dates_map.items() if '탑차' in v)

            for i, row_data in enumerate(current_exp_data.get('site_expense', [])):
                cat = str(row_data.get('cat', '')).upper()
                if any(k in cat for k in ['출장비', '소모품비', '식대']):
                    row_data['qty'] = f"{total_days_count:g}" if total_days_count > 0 else ""
                elif '차량유지비' in cat:
                    row_data['qty'] = f"{total_vehicle_days:g}" if total_vehicle_days > 0 else ""
                elif 'SE-175' in cat or 'SE175' in cat:
                    row_data['qty'] = f"{len(rt_dates):g}" if len(rt_dates) > 0 else ""
                # [FIX] 사전예산 단가 복사
                if 'site_expense' in planned_exp_data and i < len(planned_exp_data['site_expense']):
                    row_data['unit_price'] = planned_exp_data['site_expense'][i].get('unit_price', '')
                    
            for i, row_data in enumerate(current_exp_data.get('depreciation', [])):
                item = row_data.get('item', '')
                
                updated_days = False
                if '스타렉스' in item:
                    row_data['days'] = f"{starex_days:g}" if starex_days > 0 else ""
                    if starex_days > 0: updated_days = True
                elif '탑차' in item:
                    row_data['days'] = f"{toptruck_days:g}" if toptruck_days > 0 else ""
                    if toptruck_days > 0: updated_days = True
                else:
                    item_upper_clean = str(item).upper().replace(' ', '')
                    if 'SCANNER(MANUAL)' in item_upper_clean:
                        row_data['days'] = f"{len(paut_manual_scanner_dates):g}" if len(paut_manual_scanner_dates) > 0 else ""
                        if len(paut_manual_scanner_dates) > 0: updated_days = True
                    elif 'SCANNER(COBRA)' in item_upper_clean:
                        row_data['days'] = f"{len(paut_cobra_scanner_dates):g}" if len(paut_cobra_scanner_dates) > 0 else ""
                        if len(paut_cobra_scanner_dates) > 0: updated_days = True
                    elif 'PAUT' in item_upper_clean:
                        row_data['days'] = f"{len(paut_dates):g}" if len(paut_dates) > 0 else ""
                        if len(paut_dates) > 0: updated_days = True
                    elif 'YOKE' in str(item).upper() or ('MT' in str(item).upper() and 'PAUT' not in str(item).upper()):
                        row_data['days'] = f"{len(mt_dates):g}" if len(mt_dates) > 0 else ""
                        if len(mt_dates) > 0: updated_days = True
                    else:
                        item_upper = str(item).upper().strip()
                        matched_days = 0
                        for e_name, e_dates in equip_dates_map.items():
                            if e_name.upper().strip() in item_upper or item_upper in e_name.upper().strip():
                                matched_days = max(matched_days, len(e_dates))
                        if matched_days > 0:
                            row_data['days'] = f"{matched_days:g}"
                            updated_days = True

                # [FIX] 사전예산 단가 복사 및 수량 처리
                if 'depreciation' in planned_exp_data and i < len(planned_exp_data['depreciation']):
                    row_data['rate'] = planned_exp_data['depreciation'][i].get('rate', '')
                    
                    if updated_days:
                        # 일수가 자동 합산된 항목은 중복곱셈을 막기 위해 수량을 무조건 1로 강제
                        row_data['qty'] = '1'
                    else:
                        # 자동 합산되지 않은 항목은 빈칸일 때만 사전예산 수량 복사
                        current_qty = str(row_data.get('qty', '')).strip()
                        planned_qty = str(planned_exp_data['depreciation'][i].get('qty', '')).strip()
                        if not current_qty:
                            row_data['qty'] = planned_qty if planned_qty else '1'
                else:
                    if updated_days or not str(row_data.get('qty', '')).strip():
                        row_data['qty'] = '1'
            
            self.actual_expense_detail_widget.set_data(current_exp_data)

        # KPI & Diff Update
        self._update_budget_kpis()
        self.root.update_idletasks()

        # [NEW] Calculate profit/margin for the summary message
        try:
            a_rev = float(self.ent_budget_actual_unit_price.get().replace(',', '') or 0)
            if a_rev == 0:
                a_rev = float(self.ent_budget_actual_revenue.get().replace(',', '') or 0)
            a_lab = float(self.ent_budget_actual_labor.get().replace(',', '') or 0)
            a_mat = float(self.ent_budget_actual_material.get().replace(',', '') or 0)
            a_exp = float(self.ent_budget_actual_expense.get().replace(',', '') or 0)
            a_out = float(self.ent_budget_actual_outsource.get().replace(',', '') or 0)
            a_cost = a_lab + a_mat + a_exp + a_out
            a_profit = a_rev - a_cost
            a_margin = (a_profit / a_rev * 100) if a_rev > 0 else 0.0
            
            if hasattr(self, 'ent_budget_actual_profit'):
                self.ent_budget_actual_profit.delete(0, tk.END)
                self.ent_budget_actual_profit.insert(0, f"{a_profit:,.0f} ({a_margin:.1f}%)")

            profit_str = f"\n\n▶ 실적 집계 결과\n- 도급액(실적): {a_rev:,.0f}원\n- 사후원가: {a_cost:,.0f}원\n- 현재이익: {a_profit:,.0f}원 ({a_margin:.1f}%)"
        except:
            profit_str = ""

        messagebox.showinfo("조회 완료", 
                            f"'{site}' 현장의 실적 데이터를 집계하여 [사후원가(실적)] 칸에 표시하였습니다.\n"
                            f"사전예산은 방어되어 변경되지 않았습니다.\n"
                            f"(기간: {start_ts.strftime('%Y-%m-%d')} ~ {end_ts.strftime('%Y-%m-%d')}){profit_str}")

    def save_budget_entry(self):
        """Save or update budget entry"""
        site = self.cb_budget_site.get().strip()
        if not site:
            messagebox.showwarning("입력 오류", "현장명을 입력해주세요.")
            return
            
        try:
            def _get(attr):
                try: 
                    w = getattr(self, attr, None)
                    if w: return float(str(w.get()).replace(',', '') or 0)
                    return 0.0
                except: return 0.0

            rev = _get('ent_budget_revenue')
            unit_price = _get('ent_budget_unit_price')
            lab = _get('ent_budget_labor')
            mat = _get('ent_budget_material')
            exp = _get('ent_budget_expense')
            out = _get('ent_budget_outsource')
            profit = rev - (lab + mat + exp + out)
            note = self.ent_budget_note.get().strip() if hasattr(self, 'ent_budget_note') else ""
            period = getattr(self, 'ent_budget_period').get().strip() if hasattr(self, 'ent_budget_period') else ""
            
            a_rev = _get('ent_budget_actual_revenue')
            a_unit = _get('ent_budget_actual_unit_price')
            a_lab = _get('ent_budget_actual_labor')
            a_mat = _get('ent_budget_actual_material')
            a_exp = _get('ent_budget_actual_expense')
            a_out = _get('ent_budget_actual_outsource')
            a_prof = _get('ent_budget_actual_profit')
            a_note = getattr(self, 'ent_budget_actual_note').get().strip() if hasattr(self, 'ent_budget_actual_note') else ""
            a_period = getattr(self, 'ent_budget_actual_period').get().strip() if hasattr(self, 'ent_budget_actual_period') else ""
        except ValueError:
            messagebox.showerror("입력 오류", "금액은 숫자여야 합니다.")
            return

        new_data = {
            'Site': site,
            'Period': period,
            'Revenue': rev,
            'UnitPrice': unit_price,
            'LaborCost': lab,
            'MaterialCost': mat,
            'Expense': exp,
            'OutsourceCost': out,
            'Profit': profit,
            'Note': note,
            'Actual_Period': a_period,
            'Actual_Revenue': a_rev, 'Actual_UnitPrice': a_unit, 'Actual_LaborCost': a_lab,
            'Actual_MaterialCost': a_mat, 'Actual_Expense': a_exp, 'Actual_OutsourceCost': a_out,
            'Actual_Profit': a_prof, 'Actual_Note': a_note,
            'LaborDetail': json.dumps(self.labor_detail_widget.get_data()) if hasattr(self, 'labor_detail_widget') else "{}",
            'MaterialDetail': json.dumps(self.material_detail_widget.get_data()) if hasattr(self, 'material_detail_widget') else "{}",
            'ExpenseDetail': json.dumps(self.expense_detail_widget.get_data()) if hasattr(self, 'expense_detail_widget') else "{}",
            'Actual_LaborDetail': json.dumps(self.actual_labor_detail_widget.get_data()) if hasattr(self, 'actual_labor_detail_widget') else "{}",
            'Actual_MaterialDetail': json.dumps(self.actual_material_detail_widget.get_data()) if hasattr(self, 'actual_material_detail_widget') else "{}",
            'Actual_ExpenseDetail': json.dumps(self.actual_expense_detail_widget.get_data()) if hasattr(self, 'actual_expense_detail_widget') else "{}"
        }

        # [NEW] Robust Sanitization: Ensure NO NaN or empty strings ever reach the float64 columns
        # This fixes TypeError: Invalid value '' for dtype 'float64' permanently
        numeric_keys = {'Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit'}
        new_data_sanitized = {}
        for k, v in new_data.items():
            if k in numeric_keys:
                try:
                    # Coerce everything to float for numeric columns
                    if pd.isna(v) or str(v).strip().lower() in ('nan', 'none', ''):
                        new_data_sanitized[k] = 0.0
                    else:
                        new_data_sanitized[k] = float(v)
                except:
                    new_data_sanitized[k] = 0.0
            else:
                # String columns (Note, Detail JSONs, Site)
                if pd.isna(v) or str(v).lower() == 'nan':
                    new_data_sanitized[k] = ""
                else:
                    new_data_sanitized[k] = str(v)


        if not self.budget_df.empty and site in self.budget_df['Site'].values:
            # [FIX] Use explicit .loc[index, key] to ensure correct row update
            idx = self.budget_df[self.budget_df['Site'] == site].index[0]
            for key, val in new_data_sanitized.items():
                self.budget_df.loc[idx, key] = val
        else:
            # Create a properly typed DataFrame for concatenation to avoid dtype clashes
            row_df = pd.DataFrame([new_data_sanitized])
            self.budget_df = pd.concat([self.budget_df, row_df], ignore_index=True)
            
        if self.save_data():
            messagebox.showinfo("성공", f"'{site}' 현장의 예산 정보가 저장/수정되었습니다.")
            
            # [NEW] Refresh UI to show the updated data in comparison rows immediately
            self.update_budget_view()
            
            if 'budget' in self.detached_windows: self._refresh_detached_budget()
            # [NEW] 저장 후 현장 목록 갱신 (새로 기입한 현장명 반영)
            self.refresh_inquiry_filters()
            # self.clear_budget_form() # [UX] Don't clear immediately so user can verify

    def _toggle_hidden_site(self, site, hide=True):
        """현장명을 드롭다운 목록에서 숨기거나 다시 표시한다."""
        if not hasattr(self, 'hidden_sites'):
            self.hidden_sites = []
        if hide:
            if site not in self.hidden_sites:
                self.hidden_sites.append(site)
            messagebox.showinfo("숨기기 완료", f"'{site}' 현장이 목록에서 숨겨졌습니다.\n다시 표시하려면 현장명 입력칸을 우클릭하세요.")
        else:
            if site in self.hidden_sites:
                self.hidden_sites.remove(site)
            messagebox.showinfo("표시 완료", f"'{site}' 현장이 목록에 다시 표시됩니다.")
        # config 저장 및 목록 갱신
        self.save_tab_config()
        self.refresh_inquiry_filters()

    def _show_all_hidden_sites(self):
        """숨긴 현장명을 모두 다시 표시한다."""
        if not getattr(self, 'hidden_sites', []):
            messagebox.showinfo("알림", "숨겨진 현장이 없습니다.")
            return
        count = len(self.hidden_sites)
        names = '\n'.join(f'  • {s}' for s in self.hidden_sites)
        confirmed = messagebox.askyesno(
            "모두 표시",
            f"현재 숨겨진 현장 {count}개:\n{names}\n\n모두 다시 표시하시겠습니까?"
        )
        if confirmed:
            self.hidden_sites.clear()
            self.save_tab_config()
            self.refresh_inquiry_filters()
            messagebox.showinfo("완료", "모든 현장이 다시 표시됩니다.")

    def delete_budget_entry(self):
        """Delete selected budget entry"""
        site = self.cb_budget_site.get().strip()
        if not site:
            messagebox.showwarning("삭제 오류", "삭제할 현장명을 입력해주세요.")
            return
            
        if messagebox.askyesno("삭제 확인", f"'{site}' 현장의 예산 정보를 삭제하시겠습니까?"):
            self.budget_df = self.budget_df[self.budget_df['Site'] != site]
            if self.save_data():
                messagebox.showinfo("성공", "삭제되었습니다.")
                self.update_budget_view()
                if 'budget' in self.detached_windows: self._refresh_detached_budget()
                # [NEW] 삭제 후 현장 목록 갱신
                self.refresh_inquiry_filters()
                self.clear_budget_form()


    def clear_budget_form(self):
        """Reset budget form fields while maintaining site selection context"""
        # self.cb_budget_site.set('')  <- Removed to preserve 'appearance' as per user request
        if hasattr(self, 'budget_widgets'):
            for k, w in self.budget_widgets.items():
                if k == "cb_budget_site": continue
                if hasattr(w, 'cget'):
                    st = w.cget('state')
                    if st == 'readonly': w.config(state='normal')
                    w.delete(0, 'end')
                    if st == 'readonly': w.config(state='readonly')
                else:
                    try: w.delete(0, 'end')
                    except: pass
        # [NEW] Reset detail widgets
        if hasattr(self, 'labor_detail_widget'): self.labor_detail_widget.reset()
        if hasattr(self, 'material_detail_widget'):
            self.material_detail_widget.reset()
        if hasattr(self, 'expense_detail_widget'):
            self.expense_detail_widget.reset()
            
        # [STABILITY] Reset KPI summary
        self._update_budget_kpis()

    def on_budget_tree_select(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import on_budget_tree_select_impl
        return on_budget_tree_select_impl(self, *args, **kwargs)

    def _ensure_sash_visible(self):
        """No-op as details panel is removed"""
        pass
    
    def _ensure_daily_usage_sash_visibility(self):
        """Ensure sash position is visible and properly sized using ratio"""
        try:
            if not hasattr(self, 'daily_usage_paned'):
                return
                
            # [STABILITY] Skip if we are still in the initial loading/restoration phase
            # load_tab_config will handle the initial placement.
            if not getattr(self, 'is_ready', False):
                return

            # Force multiple updates to get correct dimensions
            for _ in range(3):
                self.daily_usage_paned.update_idletasks()
            
            # Get current paned window dimensions
            total_h = self.daily_usage_paned.winfo_height()
            
            if total_h <= 0:
                print("PanedWindow height is 0, retrying...")
                self.root.after(100, self._ensure_daily_usage_sash_visibility)
                return
            
            # If locked, prioritize absolute position if valid
            if hasattr(self, 'daily_usage_sash_locked') and self.daily_usage_sash_locked:
                if hasattr(self, 'tab_config') and 'daily_usage_sash_pos' in self.tab_config:
                    target_pos = int(self.tab_config['daily_usage_sash_pos'])
                    # Ensure within bounds
                    if 50 < target_pos < total_h - 50:
                        getattr(self.daily_usage_paned, "sashpos", lambda *args: 500)(0, target_pos)
                        print(f"Restored locked absolute position: {target_pos}")
                        return

            # Try to restore from saved ratio first
            # [UX FIX] 작업자 박스 하단이 가려지지 않도록 상단 영역 기본 비율 상향
            target_ratio = 0.45  # Default ratio
            if hasattr(self, 'tab_config') and 'daily_usage_sash_ratio' in self.tab_config:
                target_ratio = self.tab_config['daily_usage_sash_ratio']
            
            # Calculate target position
            target_pos = int(total_h * target_ratio)
            
            # Apply relaxed bounds + practical minimum top area (Reduced from 460 to 100 to allow user custom sizing)
            min_pos = max(int(total_h * 0.1), 100)
            max_pos = int(total_h * 0.95)
            
            if target_pos < min_pos:
                target_pos = min_pos
            elif target_pos > max_pos:
                target_pos = max_pos
            
            # Set sash position
            getattr(self.daily_usage_paned, "sashpos", lambda *args: 500)(0, target_pos)
            
            # Verify position was set correctly
            actual_pos = 500
            actual_ratio = actual_pos / total_h if total_h > 0 else target_ratio
            
            # Update config
            if hasattr(self, 'is_ready') and self.is_ready:
                if not hasattr(self, 'tab_config'):
                    self.tab_config = {}
                self.tab_config['daily_usage_sash_ratio'] = actual_ratio
                self.tab_config['daily_usage_sash_pos'] = actual_pos
            
            # Update canvas scroll region
            self.root.after(100, self._ensure_canvas_scroll_region)
            
        except Exception as e:
            print(f"Error ensuring sash visibility: {e}")
            self.root.after(500, self._ensure_daily_usage_sash_visibility)
    
    
    
    def update_resolution_display(self):
        """Update the resolution display"""
        try:
            if hasattr(self, 'resolution_label'):
                # Get window dimensions
                window_width = self.root.winfo_width()
                window_height = self.root.winfo_height()
                
                # Get screen dimensions
                screen_width = self.root.winfo_screenwidth()
                screen_height = self.root.winfo_screenheight()
                
                # Get daily usage paned window dimensions
                if hasattr(self, 'daily_usage_paned'):
                    self.daily_usage_paned.update_idletasks()
                    pane_height = self.daily_usage_paned.winfo_height()
                    pane_width = self.daily_usage_paned.winfo_width()
                    
                    # Get current sash position and ratio
                    try:
                        sash_pos = 500
                        ratio = (sash_pos / pane_height * 100) if pane_height > 0 else 0
                        resolution_text = f"창: {window_width}x{window_height} | 화면: {screen_width}x{screen_height} | 패널: {pane_width}x{pane_height} | 경계: {ratio:.1f}%"
                    except:
                        resolution_text = f"창: {window_width}x{window_height} | 화면: {screen_width}x{screen_height} | 패널: {pane_width}x{pane_height}"
                else:
                    resolution_text = f"창: {window_width}x{window_height} | 화면: {screen_width}x{screen_height}"
                
                self.resolution_label.config(text=resolution_text)
                
                # Schedule next update
                self.root.after(500, self.update_resolution_display)
        except Exception as e:
            print(f"Error updating resolution display: {e}")
            # Still schedule next update even if there's an error
            if hasattr(self, 'resolution_label'):
                self.root.after(1000, self.update_resolution_display)
    
    def _restore_locked_position(self):
        """Restore sash to locked position"""
        try:
            if not hasattr(self, 'daily_usage_paned'): return
            
            if hasattr(self, 'tab_config'):
                total_h = self.daily_usage_paned.winfo_height()
                if total_h <= 0: return
                
                # Prioritize absolute position when locked
                if 'daily_usage_sash_pos' in self.tab_config:
                    pos = int(self.tab_config['daily_usage_sash_pos'])
                    # Only apply if it doesn't hide the bottom area completely
                    if 50 < pos < total_h - 50:
                        getattr(self.daily_usage_paned, "sashpos", lambda *args: 500)(0, pos)
                        return
                
                # Fallback to ratio
                if 'daily_usage_sash_ratio' in self.tab_config:
                    ratio = self.tab_config['daily_usage_sash_ratio']
                    locked_pos = int(total_h * ratio)
                    getattr(self.daily_usage_paned, "sashpos", lambda *args: 500)(0, locked_pos)
        except Exception as e:
            print(f"Error restoring locked position: {e}")
    
    def _on_daily_usage_resize(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import _on_daily_usage_resize_impl
        return _on_daily_usage_resize_impl(self, *args, **kwargs)
            # Fallback if something went wrong during resize

    def toggle_sash_lock(self):
        """Toggle sash lock state with auto-fit logic"""
        try:
            self.daily_usage_sash_locked = not self.daily_usage_sash_locked
            
            if self.daily_usage_sash_locked:
                self.root.update_idletasks()
                
                total_height = self.daily_usage_paned.winfo_height()
                sash_pos = 500
                
                if not hasattr(self, 'tab_config'):
                    self.tab_config = {}
                
                # Calculate and save ratio and absolute position
                ratio = sash_pos / total_height if total_height > 0 else 0.5
                self.tab_config['daily_usage_sash_ratio'] = ratio
                self.tab_config['daily_usage_sash_pos'] = sash_pos
                self.tab_config['daily_usage_sash_locked'] = True
                
                # Save configuration immediately
                self.save_tab_config(force=True)
                
                if hasattr(self, 'btn_sash_lock'):
                    self.btn_sash_lock.config(text="🔒 경계 고정됨")
                    self.btn_sash_lock.configure(style="SashLock.TButton")
                    self.style.configure("SashLock.TButton", foreground="red")
                
                self._start_sash_monitor()
                print(f"Daily usage sash position LOCKED at height: {sash_pos}")
            else:
                # Unlock
                if hasattr(self, 'btn_sash_lock'):
                    self.btn_sash_lock.config(text="🔓 경계 자유")
                    self.btn_sash_lock.configure(style="TButton")
                
                self._stop_sash_monitor()
                if hasattr(self, 'tab_config'):
                    self.tab_config['daily_usage_sash_locked'] = False
                    self.save_tab_config(force=True)
                print("Daily usage sash position UNLOCKED")
            
            # [NEW] Refresh scroll region to ensure it stops correctly at the new boundary
            self._ensure_canvas_scroll_region()
        except Exception as e:
            print(f"Error toggling sash lock: {e}")
    
    def _start_sash_monitor(self):
        """Start periodic monitoring of sash position"""
        if hasattr(self, '_sash_monitor_job'):
            self.root.after_cancel(self._sash_monitor_job)
        
        def check_sash():
            if hasattr(self, 'daily_usage_sash_locked') and self.daily_usage_sash_locked:
                self._restore_locked_position()
                # Schedule next check
                self._sash_monitor_job = self.root.after(200, check_sash)
        
        # Start monitoring
        self._sash_monitor_job = self.root.after(200, check_sash)
        print("Started sash position monitoring")
    
    def _stop_sash_monitor(self):
        """Stop periodic monitoring of sash position"""
        if hasattr(self, '_sash_monitor_job'):
            self.root.after_cancel(self._sash_monitor_job)
            delattr(self, '_sash_monitor_job')
            print("Stopped sash position monitoring")
    
    def _on_main_window_resize(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import _on_main_window_resize_impl
        return _on_main_window_resize_impl(self, *args, **kwargs)
    
    
    def show_error_dialog(self, title, message):
        """Show a custom error dialog with draggable text"""
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("600x400")
        dialog.resizable(True, True)
        
        # Make dialog modal
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (600 // 2)
        y = (dialog.winfo_screenheight() // 2) - (400 // 2)
        dialog.geometry(f"600x400+{x}+{y}")
        
        # Main frame
        main_frame = ttk.Frame(dialog)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Title label
        title_label = ttk.Label(main_frame, text=title, font=('Malgun Gothic', 12, 'bold'))
        title_label.pack(pady=(0, 10))
        
        # Text widget with scrollbar for error message
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill='both', expand=True)
        
        text_widget = tk.Text(text_frame, wrap='word', font=('Malgun Gothic', 10))
        scrollbar = ttk.Scrollbar(text_frame, orient='vertical', command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Insert error message
        text_widget.insert('1.0', message)
        text_widget.configure(state='normal')  # Allow selection and copying
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill='x', pady=(10, 0))
        
        # Copy button
        def copy_text():
            dialog.clipboard_clear()
            selected_text = text_widget.get('sel.first', 'sel.last') if text_widget.tag_ranges('sel') else text_widget.get('1.0', 'end')
            dialog.clipboard_append(selected_text)
        
        copy_btn = ttk.Button(button_frame, text="복사하기", command=copy_text)
        copy_btn.pack(side='left', padx=5)
        
        # Close button
        close_btn = ttk.Button(button_frame, text="닫기", command=dialog.destroy)
        close_btn.pack(side='right', padx=5)
        
        # Focus on text widget
        text_widget.focus_set()
        
        # Bind Escape key to close
        dialog.bind('<Escape>', lambda e: dialog.destroy())
        
        # Wait for dialog to close
        dialog.wait_window()

    def on_daily_usage_tree_select(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import on_daily_usage_tree_select_impl
        return on_daily_usage_tree_select_impl(self, *args, **kwargs)

    def delete_selected_site(self):
        """Remove the selected site from the suggestion list"""
        site = self.cb_daily_site.get().strip()
        if not site:
            messagebox.showwarning("선택 오류", "삭제할 현장명을 선택해주세요.")
            return
            
        if site in self.sites:
            if messagebox.askyesno("삭제 확인", f"'{site}' 현장명을 목록에서 삭제하시겠습니까?\n(기존 기록은 삭제되지 않습니다.)"):
                self.sites.remove(site)
                self.cb_daily_site['values'] = self.sites
                self.cb_daily_site.set('')
                self.save_tab_config()
                messagebox.showinfo("완료", "현장명이 삭제되었습니다.")
        else:
            messagebox.showinfo("알림", "목록에 없는 현장명입니다.")

    def calculate_rtk_total(self):
        """Calculate total RTK usage"""
        try:
            total = 0
            rtk_categories = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타"]
            for category in rtk_categories:
                value = self.rtk_entries[category].get()
                if value:
                    total += float(value)
            self.rtk_entries["총계"].config(state='normal')
            self.rtk_entries["총계"].delete(0, tk.END)
            # Format as integer if possible, else float
            if total == int(total):
                self.rtk_entries["총계"].insert(0, str(int(total)))
            else:
                self.rtk_entries["총계"].insert(0, f"{total:.1f}")
            self.rtk_entries["총계"].config(state='readonly')
        except ValueError:
            pass  # Ignore invalid input during typing
    
    def format_entry_with_commas(self, *args, **kwargs):
        from site_apps.central.src.utils.helpers import format_entry_with_commas_impl
        return format_entry_with_commas_impl(self, *args, **kwargs)

    def update_daily_test_fee_calc(self, event=None):
        """Auto-calculate Inspection Fee = (Amount * Unit Price) + Travel Expense"""
        try:
            def get_f(entry):
                try:
                    val = entry.get().strip().replace(',', '')
                    return float(val) if val else 0.0
                except: return 0.0

            amount = get_f(self.ent_daily_test_amount)
            price = get_f(self.ent_daily_unit_price)
            travel = get_f(self.ent_daily_travel_cost)
            
            calc_fee = (amount * price) + travel
            
            # Update the fee field with commas
            self.ent_daily_test_fee.delete(0, tk.END)
            self.ent_daily_test_fee.insert(0, f"{calc_fee:,.0f}")
        except:
            pass

    def _add_single_usage_record_logic(self, mat_id, date_val, site, auto_save=True):
        """현장 사용량 기록 1건을 처리하는 핵심 로직 (단일/일괄 공용)"""
        def to_f(ent):
            try:
                # Handle both Entry widgets and raw values
                val = ent.get().replace(',', '') if hasattr(ent, 'get') else str(ent).replace(',', '')
                return float(val) if val else 0.0
            except: return 0.0

        # 1. 작업자 목록 및 시간 데이터 구성
        workers_list = []
        worker_data_map = {}
        for i in range(1, 11):
            group = getattr(self, f'worker_group{i}', None)
            u_key = 'User' if i == 1 else f'User{i}'
            wt_key = 'WorkTime' if i == 1 else f'WorkTime{i}'
            ot_key = 'OT' if i == 1 else f'OT{i}'
            m_key = 'Meal' if i == 1 else f'Meal{i}'
            
            if group:
                name = group.get_worker().strip()
                wt = group.get_time().strip()
                ot = group.get_ot().strip()
                meal = group.get_meal().strip()
                worker_data_map[u_key] = name
                worker_data_map[wt_key] = wt
                worker_data_map[ot_key] = ot
                worker_data_map[m_key] = meal
                if name: workers_list.append(name)
            else:
                worker_data_map[u_key] = ""
                worker_data_map[wt_key] = ""
                worker_data_map[ot_key] = ""
                worker_data_map[m_key] = ""
        
        all_workers = ", ".join(workers_list)

        # 2. RTK 데이터 구성
        rtk_data = {}
        for cat, ent in self.rtk_entries.items():
            rtk_data[f'RTK_{cat}'] = to_f(ent)

        # 3. 차량 데이터 수집 (하단 고정 패널 1개 + 추가 플로팅 창)
        all_v_widgets = []
        if hasattr(self, 'fixed_vehicle_widget') and self.fixed_vehicle_widget.winfo_exists():
            all_v_widgets.append(self.fixed_vehicle_widget)
        if hasattr(self, 'vehicle_boxes'):
            for b in self.vehicle_boxes:
                if b not in all_v_widgets:
                    all_v_widgets.append(b)
            
        living_boxes = []
        for b in all_v_widgets:
            if hasattr(b, 'winfo_exists') and b.winfo_exists():
                living_boxes.append(b)
                
        # 4. 업체(Company) 데이터 수집
        company_data_list = []
        if hasattr(self, 'ndt_company_entries'):
            for company_entries in self.ndt_company_entries:
                company_code = company_entries['_company'].get().strip()
                ndt_data = {name: to_f(company_entries.get(name)) for name in self.ndt_materials_all}
                
                # [REFINEMENT] Create record if there is NDT data OR it is the first entry (guaranteed row)
                if any(v > 0 for v in ndt_data.values()) or company_entries == self.ndt_company_entries[0]:
                    company_data_list.append({
                        '회사코드': company_code,
                        'ndt_data': ndt_data,
                        'is_primary': company_entries == self.ndt_company_entries[0]
                    })
        
        if not company_data_list: # [FALLBACK] Ensure at least one row exists
             company_data_list.append({'회사코드': "", 'ndt_data': {}, 'is_primary': True})

        # 5. 공통 레코드 데이터 구성 (기본값)
        common_data = {
            'Date': date_val,
            'Site': site,
            'MaterialID': mat_id,
            '장비명': self.cb_daily_equip.get().strip(),
            '검사방법': self.cb_daily_test_method.get().strip(),
            '검사품명': self.ent_daily_inspection_item.get().strip(),
            '적용코드': self.ent_daily_applied_code.get().strip(),
            '성적서번호': self.ent_daily_report_no.get().strip(),
            '검사자': "", # Derived from workers list in output
            'Usage': to_f(self.ent_daily_test_amount),
            '검사량': to_f(self.ent_daily_test_amount),
            '단가': to_f(self.ent_daily_unit_price),
            '출장비': to_f(self.ent_daily_travel_cost),
            '업체명': self.cb_daily_company.get().strip(),
            'Unit': self.cb_daily_unit.get().strip(),
            '작업형태': self.ndt_work_time_var.get() if self.cb_daily_test_method.get().strip() in ['RT','UT','PT','PAUT'] else "",
            '조건1': self.ndt_source_var.get() if self.cb_daily_test_method.get().strip() == 'RT' else (self.ndt_pipe_var.get() if self.cb_daily_test_method.get().strip() in ['UT','PT','PAUT'] else ""),
            '조건2': self.ndt_thickness_var.get() if self.cb_daily_test_method.get().strip() in ['RT','UT'] else "",
            '제경비': getattr(self, '_last_ndt_overhead', 0) if self.cb_daily_test_method.get().strip() in ['RT','UT','PT','PAUT'] else 0,
            '기술료': getattr(self, '_last_ndt_tech', 0) if self.cb_daily_test_method.get().strip() in ['RT','UT','PT','PAUT'] else 0,
            '보정계수': getattr(self, '_last_ndt_factor', 1.0) if self.cb_daily_test_method.get().strip() in ['RT','UT','PT','PAUT'] else 1.0,
            '환산물량': getattr(self, '_last_ndt_adj_qty', 0.0) if self.cb_daily_test_method.get().strip() in ['RT','UT','PT','PAUT'] else 0.0,
            '재료비': getattr(self, '_last_ndt_mat_cost', 0) if self.cb_daily_test_method.get().strip() in ['RT','UT','PT','PAUT'] else 0,
            '인건비': getattr(self, '_last_ndt_lab_cost', 0) if self.cb_daily_test_method.get().strip() in ['RT','UT','PT','PAUT'] else 0,
            '검사구분': "ORI",
            '구분': self.ndt_loc_type_var.get().strip() if self.cb_daily_test_method.get().strip() in ['RT','UT','PT','PAUT'] else "",
            '조인트수': "",
            '불량수': "",
            '관경(Inch)': getattr(self, 'ndt_report_pipe_var', tk.StringVar(value="")).get().strip(),
        }

        # [NEW] Auto-save new unit, equipment, and site (with improved duplicate check)
        new_unit = common_data['Unit'].strip()
        if new_unit and not any(u.strip() == new_unit for u in self.daily_units):
            self.daily_units.append(new_unit)
            self.refresh_ui_for_list_change('daily_units')

        new_equip = common_data['장비명'].strip()
        if new_equip and not any(e.strip() == new_equip for e in self.equipments):
            self.equipments.append(new_equip)
            self.refresh_ui_for_list_change('equipments')

        new_site = common_data['Site'].strip()
        if new_site and not any(s.strip() == new_site for s in self.sites):
            self.sites.append(new_site)
            self.refresh_ui_for_list_change('sites')

        total_meal_worker = sum(to_f(worker_data_map.get(f'Meal{i}' if i > 1 else 'Meal', 0)) for i in range(1, 11))
        global_meal = to_f(self.ent_daily_meal_cost)
        
        # [NEW] Append REP defect trace to Note if applicable (Option A)
        base_note = self._get_merged_memo_and_note()
        rep_count = getattr(self, 'ndt_rep_joint_count_var', tk.StringVar(value="")).get().strip()
        insp_type = getattr(self, 'ndt_inspection_type_var', tk.StringVar(value="ORI")).get().strip()
        if insp_type == "ORI" and rep_count and rep_count != "0":
            base_note = (base_note + f" [불량(REP) {rep_count}개 발생]").strip()
        
        # [NEW] Collect DailyMemo
        daily_memo_text = ""
        if hasattr(self, 'main_memo_text'):
            try: daily_memo_text = self.main_memo_text.get('1.0', 'end-1c').strip()
            except: pass

        common_data.update({
            '일식': total_meal_worker if total_meal_worker > 0 else global_meal,
            '검사비': to_f(self.ent_daily_test_fee),
            'FilmCount': 0.0,
            'EntryTime': datetime.datetime.now(),
            '회사코드': "",
            'DailyMemo': daily_memo_text,
            **worker_data_map
        })

        is_ndt = self.cb_daily_test_method.get().strip() in ['RT','UT','PT','PAUT']
        record_types = []
        if is_ndt:
            ori_j = getattr(self, 'ndt_ori_joint_var', tk.StringVar(value="")).get().strip()
            rep_j = getattr(self, 'ndt_rep_joint_var', tk.StringVar(value="")).get().strip()
            if ori_j or not rep_j:
                record_types.append('ORI')
            if rep_j:
                record_types.append('REP')
        else:
            record_types.append('DEFAULT')

        # --- 차량 데이터 병합 (단일 행 저장용) ---
        merged_v_no = []
        merged_v_mileage = []
        merged_v_check = []
        merged_v_remarks = []

        for v_widget in living_boxes:
            v_data = v_widget.get_data()
            v_no = str(v_data.get('vehicle_info', '')).strip()
            v_mileage = str(v_data.get('_raw_mileage', '')).strip()  # Use raw mileage for safety
            v_remarks = str(v_data.get('remarks', '')).strip()
            
            reserved = ['vehicle_info', 'mileage', 'remarks', '_raw_mileage']
            checks_list = []
            for k, v in v_data.items():
                if k not in reserved and v:
                    checks_list.append(f"{k}:{v}")
            v_check_str = "|".join(checks_list)
            
            # 값이 하나라도 있으면 추가
            if v_no or v_mileage or v_check_str or v_remarks:
                merged_v_no.append(v_no)
                merged_v_mileage.append(v_mileage)
                merged_v_check.append(v_check_str)
                merged_v_remarks.append(v_remarks)
        
        final_v_no = " || ".join(merged_v_no)
        final_v_mileage = " || ".join(merged_v_mileage)
        final_v_check = " || ".join(merged_v_check)
        final_v_remarks = " || ".join(merged_v_remarks)

        records_to_save = []
        # 차량 개수로 행이 늘어나지 않게 max_rows 고정
        max_rows = max(len(company_data_list), 1)
        
        for r_type in record_types:
            r_common_data = common_data.copy()
            
            if is_ndt:
                r_common_data['검사구분'] = r_type
                if r_type == 'ORI':
                    r_common_data['조인트수'] = getattr(self, 'ndt_ori_joint_var', tk.StringVar(value="")).get().strip()
                    qty_str = getattr(self, 'ndt_ori_qty_var', tk.StringVar(value="")).get().strip()
                    if qty_str:
                        r_qty = float(qty_str)
                        r_common_data['Usage'] = r_qty
                        r_common_data['검사량'] = r_qty
                    
                    base_note = self._get_merged_memo_and_note()
                    rej_count = getattr(self, 'ndt_rej_joint_var', tk.StringVar(value="")).get().strip()
                    r_common_data['불량수'] = rej_count if rej_count else "0"
                    if rej_count and rej_count != "0":
                        base_note = (base_note + f" [당일 불량(REJ) {rej_count}개 발생]").strip()
                    r_common_data['Note'] = base_note
                elif r_type == 'REP':
                    r_common_data['조인트수'] = getattr(self, 'ndt_rep_joint_var', tk.StringVar(value="")).get().strip()
                    qty_str = getattr(self, 'ndt_rep_qty_var', tk.StringVar(value="")).get().strip()
                    if qty_str:
                        r_qty = float(qty_str)
                        r_common_data['Usage'] = r_qty
                        r_common_data['검사량'] = r_qty
                    
                    r_common_data['Note'] = "(재검사 기록) " + self._get_merged_memo_and_note()
                    r_common_data['출장비'] = 0.0
                    r_common_data['일식'] = 0.0
                    r_common_data['검사비'] = 0.0 
            else:
                r_common_data['검사구분'] = ""
                r_common_data['Note'] = self._get_merged_memo_and_note()

            for i in range(max_rows):
                row_record = r_common_data.copy()
                
                # --- 업체/수량 데이터 배분 ---
                if i < len(company_data_list):
                    cd = company_data_list[i]
                    row_record['회사코드'] = cd['회사코드']
                    
                    if not cd['is_primary'] or r_type == 'REP':
                        for j in range(1, 11):
                            wt_k = 'WorkTime' if j == 1 else f'WorkTime{j}'
                            ot_k = 'OT' if j == 1 else f'OT{j}'
                            row_record[wt_k] = ""
                            row_record[ot_k] = ""
                        for k in rtk_data: row_record[k] = 0.0
                    else:
                        for k, v in rtk_data.items(): row_record[k] = v

                    if not cd['is_primary']:
                        for k in ['Usage', '검사량', '단가', '출장비', '검사비', '일식', 'FilmCount']:
                            row_record[k] = 0.0

                    for k, v in cd['ndt_data'].items():
                        row_record[f'NDT_{k}'] = v if r_type != 'REP' else 0.0
                else:
                    for k in ['Usage', '검사량', '단가', '출장비', '검사비', '일식', 'FilmCount']: row_record[k] = 0.0
                    for k in rtk_data: row_record[k] = 0.0
                    for j in range(1, 11):
                        row_record['WorkTime' if j==1 else f'WorkTime{j}'] = ""
                        row_record['OT' if j==1 else f'OT{j}'] = ""
                    for name in self.ndt_materials_all:
                        row_record[f'NDT_{name}'] = 0.0

                # --- 차량 데이터 배분 ---
                if i == 0 and r_type != 'REP':
                    row_record['차량번호'] = final_v_no
                    row_record['주행거리'] = final_v_mileage
                    row_record['차량점검'] = final_v_check
                    row_record['차량비고'] = final_v_remarks
                else:
                    row_record['차량번호'] = ""
                    row_record['주행거리'] = ""
                    row_record['차량점검'] = ""
                    row_record['차량비고'] = ""
                
                records_to_save.append(row_record)

        # 7. 재고 트랜잭션 처리 (원본 로직 유지 - 출고는 실제 물량만큼 한 번만 발생)
        # [NEW] PAUT 및 장비류는 재고 차감에서 제외 (사용자 요청)
        method_raw = common_data.get('검사방법', '')
        method = str(method_raw).strip().upper()
        is_excluded = (method == 'PAUT')
        
        if not is_excluded:
            # 7-1. NDT 약품 차감 (세척제, 현상제 등)
            for cd in company_data_list:
                 if cd.get('ndt_data') and any(v > 0 for v in cd['ndt_data'].values()):
                      self._auto_reconcile_and_register_ndt(date_val, site, cd['ndt_data'], all_workers, cd['회사코드'])
            
            # 7-2. 메인 자재 차감 (RT 필름 등)
            if mat_id and common_data['검사량'] > 0:
                mat_display = self.get_material_display_name(mat_id)
                # [FIX] 소모품인 경우에만 재고 차감 (장비류 제외)
                if self._is_consumable_material(mat_display, method):
                    self._create_manual_stock_transaction(date_val, mat_id, 'OUT', common_data['검사량'], site, all_workers, f"{site} 현장 사용 (자동 차감)")
                else:
                    print(f"[DEBUG] Skipping stock deduction for equipment/excluded: {mat_display}")

        # 8. 데이터프레임 업데이트
        if records_to_save:
            self.daily_usage_df = pd.concat([self.daily_usage_df, pd.DataFrame(records_to_save)], ignore_index=True)
            
            # 리스트 명단 자동 추가 (Master List 관리)
            any_list_changed = False
            if site and site not in self.sites:
                self.sites.append(site); self.sites.sort(); any_list_changed = True
            for w in workers_list:
                if w and w not in self.users:
                    self.users.append(w); self.users.sort(); any_list_changed = True
            
            # 모든 생성된 레코드에서 차량번호 수집 및 추가
            current_bases = {v.split('(')[0].strip(): v for v in self.vehicles}
            for r in records_to_save:
                v_no_raw = r.get('차량번호', '').strip()
                if v_no_raw and v_no_raw.lower() not in ['nan', 'none']:
                    v_list = [v.strip() for v in v_no_raw.replace('||', ',').split(',') if v.strip()]
                    for v_no in v_list:
                        base_no = v_no.split('(')[0].strip()
                        if base_no not in current_bases:
                            if any(t in base_no for t in ['81두1580', '89보4028', '90너4889']):
                                new_v = f"{base_no} (탑차)"
                            else:
                                new_v = f"{base_no} (스타렉스)"
                            self.vehicles.append(new_v)
                            current_bases[base_no] = new_v
                            any_list_changed = True
            
            if any_list_changed:
                self.vehicles.sort()
                self.refresh_ui_for_list_change('vehicles') # Update UI comboboxes immediately
                self.save_tab_config()

        if auto_save:
            res = self.save_data()
            return len(records_to_save) if res else 0
        return len(records_to_save)

    def add_ndt_company_section(self):
        """회사별 NDT 입력 섹션 추가"""
        if not hasattr(self, 'ndt_company_entries'):
            self.ndt_company_entries = []
        
        company_idx = len(self.ndt_company_entries)
        ndt_materials = self.ndt_materials_all


        
        # Company frame (Standard frame to save space, no title border)
        company_frame = ttk.Frame(self.ndt_company_container)
        company_frame.pack(fill='x', padx=2, pady=2)
        
        # Single Grid frame for everything to save vertical space
        grid_frame = ttk.Frame(company_frame)
        for c in range(8): grid_frame.columnconfigure(c, weight=1, uniform="ndt_rtk")
        grid_frame.pack(fill='x', padx=2, pady=2)
        
        # Company code selector (Index 0)
        ttk.Label(grid_frame, text="회사코드:", font=('Arial', 8, 'bold')).grid(row=0, column=0, padx=1, pady=1, sticky='e')
        cb_co = ttk.Combobox(grid_frame, width=8, values=getattr(self, 'co_code_list', []))
        cb_co.grid(row=0, column=1, padx=1, pady=1, sticky='ew')
        cb_co.set('')  # Default empty
        
        entries = {'_company': cb_co}  # Store company combobox
        
        # NDT entries grid (Index 1 to 7)
        for i, mat in enumerate(ndt_materials):
            idx = i + 1
            r = idx // 4
            c = (idx % 4) * 2
            ttk.Label(grid_frame, text=f"{mat}:", font=('Arial', 8)).grid(row=r, column=c, padx=1, pady=1, sticky='e')
            e = ttk.Entry(grid_frame, width=6)
            e.grid(row=r, column=c+1, padx=1, pady=1, sticky='ew')
            entries[mat] = e
        
        self.ndt_company_entries.append(entries)
    
    def remove_last_ndt_company(self):
        """마지막 회사 섹션 삭제"""
        if hasattr(self, 'ndt_company_entries') and len(self.ndt_company_entries) > 1:
            # Remove last entry dict
            self.ndt_company_entries.pop()
            # Destroy last widget in container
            widgets = self.ndt_company_container.winfo_children()
            if widgets:
                widgets[-1].destroy()

    def sync_worker_times(self, *args, **kwargs):
        from site_apps.central.src.models.worker_model import sync_worker_times_impl
        return sync_worker_times_impl(self, *args, **kwargs)
    def _load_ndt_product_map(self):
        """config 에서 NDT 약품 → 실제 DB 품목명 매핑 불러오기"""
        try:
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                return cfg.get('ndt_product_map', {})
        except Exception:
            pass
        return {}

    def _save_ndt_product_map(self, map_data):
        """NDT 약품 -> 실제 DB 품목명 매핑 저장"""
        try:
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
            else:
                cfg = {}
            cfg['ndt_product_map'] = map_data
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, indent=4, ensure_ascii=False)
            return True
        except Exception:
            return False

    def open_ndt_product_map_dialog(self):
        """NDT 약품 -> 실제 DB 품목명 매핑 설정 다이얼로그"""
        dlg = tk.Toplevel(self.root)
        dlg.title("NDT 약품-품목 매핑 설정")
        dlg.geometry("500x600")
        dlg.transient(self.root)
        dlg.grab_set()

        main_frame = ttk.Frame(dlg, padding=20)
        main_frame.pack(fill='both', expand=True)

        ttk.Label(main_frame, text="현장 입력 약품명", font=('Arial', 10, 'bold')).grid(row=0, column=0, pady=10)
        ttk.Label(main_frame, text="창고 재고 품목 (매핑)", font=('Arial', 10, 'bold')).grid(row=0, column=1, pady=10)

        ndt_materials = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
        current_map = self._load_ndt_product_map()
        
        material_options = []
        for _, row in self.materials_df.iterrows():
            disp = self.get_material_display_name(row['MaterialID'])
            material_options.append(disp)
        material_options.sort()

        combos = {}
        for i, mat in enumerate(ndt_materials):
            ttk.Label(main_frame, text=mat).grid(row=i+1, column=0, padx=5, pady=5, sticky='w')
            cb = ttk.Combobox(main_frame, values=material_options, width=40)
            cb.grid(row=i+1, column=1, padx=5, pady=5, sticky='ew')
            current_id = current_map.get(mat, "")
            if current_id:
                cb.set(self.get_material_display_name(current_id))
            combos[mat] = cb

        def _save():
            new_map = {}
            for mat, cb in combos.items():
                disp = cb.get().strip()
                if disp:
                    for _, row in self.materials_df.iterrows():
                        if self.get_material_display_name(row['MaterialID']) == disp:
                            new_map[mat] = row['MaterialID']
                            break
            if self._save_ndt_product_map(new_map):
                messagebox.showinfo("성공", "매핑 설정이 저장되었습니다.")
                dlg.destroy()
            else:
                messagebox.showerror("오류", "설정을 저장하지 못했습니다.")

        def _clear():
            for cb in combos.values():
                cb.set('')

        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=len(ndt_materials)+1, column=0, columnspan=2, pady=20)

        ttk.Button(btn_frame, text="저장", command=_save, width=10).pack(side='left', padx=8)
        ttk.Button(btn_frame, text="전체 초기화", command=_clear, width=10).pack(side='left', padx=8)
        ttk.Button(btn_frame, text="닫기", command=dlg.destroy, width=10).pack(side='left', padx=8)

    def open_equipment_management_dialog(self):
        """Dialog to add, edit or delete equipment names in the list"""
        dlg = tk.Toplevel(self.root)
        dlg.title("장비명 목록 관리")
        dlg.geometry("900x700") # Increased size
        dlg.transient(self.root)
        dlg.grab_set()

        main_frame = ttk.Frame(dlg, padding=20)
        main_frame.pack(fill='both', expand=True)

        ttk.Label(main_frame, text="사용자 정의 장비 목록", font=('Malgun Gothic', 12, 'bold')).pack(pady=(0, 5))
        ttk.Label(main_frame, text="(드롭다운 목록에 표시될 기본 장비 명단입니다)", font=('Malgun Gothic', 9)).pack(pady=(0, 10))
        
        # Treeview to show equipment list
        list_container = ttk.Frame(main_frame)
        list_container.pack(fill='both', expand=True)

        # Left side: Current Equipment List
        left_frame = ttk.LabelFrame(list_container, text="현재 장비 목록", padding=5)
        left_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))
        
        columns = ('name',)
        self.equip_tree = ttk.Treeview(left_frame, columns=columns, show='headings', height=15)
        self.equip_tree.heading('name', text='장비명')
        self.equip_tree.column('name', width=350) # Increased width
        
        scrollbar_l_y = ttk.Scrollbar(left_frame, orient="vertical", command=self.equip_tree.yview)
        scrollbar_l_x = ttk.Scrollbar(left_frame, orient="horizontal", command=self.equip_tree.xview)
        self.equip_tree.configure(yscrollcommand=scrollbar_l_y.set, xscrollcommand=scrollbar_l_x.set)
        
        self.equip_tree.pack(side='top', fill='both', expand=True)
        scrollbar_l_y.pack(side='right', fill='y', in_=left_frame)
        scrollbar_l_x.pack(side='bottom', fill='x')

        # Right side: Materials from Stock Status
        right_frame = ttk.LabelFrame(list_container, text="재고현황 품목 (가져오기)", padding=5)
        right_frame.pack(side='right', fill='both', expand=True, padx=(5, 0))

        # Search Bar for Right List
        search_frame = ttk.Frame(right_frame)
        search_frame.pack(fill='x', pady=(0, 5))
        ttk.Label(search_frame, text="🔍 검색:").pack(side='left', padx=2)
        search_var = tk.StringVar()
        search_ent = ttk.Entry(search_frame, textvariable=search_var)
        search_ent.pack(side='left', fill='x', expand=True, padx=2)

        columns_r = ('name', 'sn', 'model')
        self.mat_source_tree = ttk.Treeview(right_frame, columns=columns_r, show='headings', height=15)
        self.mat_source_tree.heading('name', text='품목명')
        self.mat_source_tree.heading('sn', text='SN')
        self.mat_source_tree.heading('model', text='모델명')
        self.mat_source_tree.column('name', width=150)
        self.mat_source_tree.column('sn', width=120)
        self.mat_source_tree.column('model', width=150)

        scrollbar_r_y = ttk.Scrollbar(right_frame, orient="vertical", command=self.mat_source_tree.yview)
        scrollbar_r_x = ttk.Scrollbar(right_frame, orient="horizontal", command=self.mat_source_tree.xview)
        self.mat_source_tree.configure(yscrollcommand=scrollbar_r_y.set, xscrollcommand=scrollbar_r_x.set)
        
        self.mat_source_tree.pack(side='top', fill='both', expand=True)
        scrollbar_r_y.pack(side='right', fill='y', in_=right_frame)
        scrollbar_r_x.pack(side='bottom', fill='x')

        def refresh_tree():
            for item in self.equip_tree.get_children():
                self.equip_tree.delete(item)
            for eq in sorted(self.equipments):
                self.equip_tree.insert('', tk.END, values=(eq,))

        def refresh_mat_tree(*args):
            query = search_var.get().strip().lower()
            for item in self.mat_source_tree.get_children():
                self.mat_source_tree.delete(item)
            
            if not self.materials_df.empty:
                # Group by unique combinations
                unique_items = self.materials_df[['품목명', 'SN', '모델명']].drop_duplicates()
                for _, row in unique_items.iterrows():
                    name = str(row.get('품목명', '')).strip()
                    sn = str(row.get('SN', '')).strip()
                    model = str(row.get('모델명', '')).strip()
                    
                    if name == 'nan': name = ''
                    if sn == 'nan': sn = ''
                    if model == 'nan': model = ''
                    
                    if name or sn or model:
                        # Apply filter
                        if not query or query in name.lower() or query in sn.lower() or query in model.lower():
                            self.mat_source_tree.insert('', tk.END, values=(name, sn, model))

        search_var.trace_add("write", refresh_mat_tree)
        refresh_tree()
        refresh_mat_tree()

        # Action Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=10)

        def add_from_stock():
            selection = self.mat_source_tree.selection()
            if not selection:
                messagebox.showwarning("선택 오류", "가져올 품목을 오른쪽 목록에서 선택해주세요.")
                return
            
            added_count = 0
            for sel in selection:
                vals = self.mat_source_tree.item(sel)['values']
                name = str(vals[0]).strip()
                sn = str(vals[1]).strip()
                model = str(vals[2]).strip()
                
                # Cleanup 'nan' strings
                if name.lower() == 'nan': name = ''
                if sn.lower() == 'nan': sn = ''
                if model.lower() == 'nan': model = ''

                # Create a display string for the equipment list
                # Format: [Model] Name (SN)
                display_parts = []
                
                # Logic: If model exists, add it. If name exists and is different from model, add it.
                if model:
                    display_parts.append(model)
                if name and name != model:
                    display_parts.append(name)
                if sn:
                    display_parts.append(f"({sn})")
                
                display_name = " ".join(display_parts).strip()
                
                if display_name and display_name not in self.equipments:
                    self.equipments.append(display_name)
                    added_count += 1
            
            if added_count > 0:
                refresh_tree()
                self.save_tab_config()
                self.update_material_combo()
                messagebox.showinfo("성공", f"{added_count}개의 품목을 장비 목록에 추가했습니다.")

        def add_equip():
            def save_new():
                name = entry.get().strip()
                if name:
                    if name not in self.equipments:
                        self.equipments.append(name)
                        refresh_tree()
                        refresh_mat_tree()
                        self.save_tab_config()
                        self.update_material_combo()
                        add_win.destroy()
                    else:
                        messagebox.showinfo("중복", "이미 존재하는 장비명입니다.")
                else:
                    messagebox.showwarning("입력 오류", "장비명을 입력해주세요.")

            add_win = tk.Toplevel(dlg)
            add_win.title("장비 추가")
            add_win.geometry("600x150")
            add_win.transient(dlg)
            add_win.grab_set()
            
            ttk.Label(add_win, text="추가할 장비명을 입력하세요:", font=('Malgun Gothic', 10)).pack(pady=10)
            entry = ttk.Entry(add_win, width=70)
            entry.pack(padx=20, pady=5)
            entry.focus_set()
            
            btn_f = ttk.Frame(add_win)
            btn_f.pack(pady=10)
            ttk.Button(btn_f, text="확인", command=save_new, width=10).pack(side='left', padx=5)
            ttk.Button(btn_f, text="취소", command=add_win.destroy, width=10).pack(side='left', padx=5)
            
            add_win.bind('<Return>', lambda e: save_new())

        def delete_equip():
            selection = self.equip_tree.selection()
            if not selection:
                messagebox.showwarning("선택 오류", "삭제할 항목을 선택해주세요.")
                return
            
            count = len(selection)
            if messagebox.askyesno("삭제 확인", f"선택한 {count}개의 장비를 목록에서 삭제하시겠습니까?"):
                for sel in selection:
                    name = self.equip_tree.item(sel)['values'][0]
                    if name in self.equipments:
                        self.equipments.remove(name)
                
                refresh_tree()
                refresh_mat_tree()
                self.save_tab_config()
                self.update_material_combo()

        def edit_equip():
            selection = self.equip_tree.selection()
            if not selection:
                messagebox.showwarning("선택 오류", "수정할 항목을 선택해주세요.")
                return
            
            old_name = self.equip_tree.item(selection[0])['values'][0]
            
            def save_edit():
                new_name = entry.get().strip()
                if new_name and new_name != old_name:
                    if old_name in self.equipments:
                        idx = self.equipments.index(old_name)
                        self.equipments[idx] = new_name
                        refresh_tree()
                        refresh_mat_tree()
                        self.save_tab_config()
                        self.update_material_combo()
                        edit_win.destroy()
                elif new_name == old_name:
                    edit_win.destroy()
                else:
                    messagebox.showwarning("입력 오류", "장비명을 입력해주세요.")

            edit_win = tk.Toplevel(dlg)
            edit_win.title("장비명 수정")
            edit_win.geometry("600x150")
            edit_win.transient(dlg)
            edit_win.grab_set()
            
            ttk.Label(edit_win, text="장비명을 수정하세요:", font=('Malgun Gothic', 10)).pack(pady=10)
            entry = ttk.Entry(edit_win, width=70)
            entry.pack(padx=20, pady=5)
            entry.insert(0, old_name)
            entry.selection_range(0, tk.END)
            entry.focus_set()
            
            btn_f = ttk.Frame(edit_win)
            btn_f.pack(pady=10)
            ttk.Button(btn_f, text="저장", command=save_edit, width=10).pack(side='left', padx=5)
            ttk.Button(btn_f, text="취소", command=edit_win.destroy, width=10).pack(side='left', padx=5)
            
            edit_win.bind('<Return>', lambda e: save_edit())

        ttk.Button(btn_frame, text="◀ 가져오기", command=add_from_stock).pack(side='left', padx=5, expand=True, fill='x')
        ttk.Button(btn_frame, text="➕ 직접추가", command=add_equip).pack(side='left', padx=5, expand=True, fill='x')
        ttk.Button(btn_frame, text="✏️ 수정", command=edit_equip).pack(side='left', padx=5, expand=True, fill='x')
        ttk.Button(btn_frame, text="❌ 삭제", command=delete_equip).pack(side='left', padx=5, expand=True, fill='x')

        ttk.Button(main_frame, text="닫기", command=dlg.destroy).pack(pady=10)

    def open_equipment_search_dialog(self):
        """Dialog to search equipment from master and select it for the form"""
        dlg = tk.Toplevel(self.root)
        dlg.title("장비 검색 및 선택")
        dlg.geometry("700x600")
        dlg.transient(self.root)
        dlg.grab_set()

        main_frame = ttk.Frame(dlg, padding=20)
        main_frame.pack(fill='both', expand=True)

        ttk.Label(main_frame, text="재고현황에서 장비 검색", font=('Malgun Gothic', 12, 'bold')).pack(pady=(0, 10))
        
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill='x', pady=5)
        ttk.Label(search_frame, text="🔍 검색어:").pack(side='left', padx=5)
        search_var = tk.StringVar()
        search_ent = ttk.Entry(search_frame, textvariable=search_var)
        search_ent.pack(side='left', fill='x', expand=True, padx=5)
        search_ent.focus_set()

        # Treeview for results
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill='both', expand=True)
        
        columns = ('name', 'sn', 'model')
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        tree.heading('name', text='품목명')
        tree.heading('sn', text='SN')
        tree.heading('model', text='모델명')
        tree.column('name', width=200)
        tree.column('sn', width=150)
        tree.column('model', width=200)
        
        sb_y = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        sb_x = ttk.Scrollbar(main_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        
        tree.pack(side='left', fill='both', expand=True)
        sb_y.pack(side='right', fill='y')
        sb_x.pack(fill='x')

        def refresh_list(*args):
            query = search_var.get().strip().lower()
            for item in tree.get_children():
                tree.delete(item)
            
            if not self.materials_df.empty:
                # [STRICT] Filter for Active and Non-Consumable (Equipment) items only
                active_mask = pd.to_numeric(self.materials_df.get('Active', 1), errors='coerce').fillna(1) > 0.5
                eq_df = self.materials_df[active_mask]
                
                # Further filter: Exclude consumables
                equipment_df = eq_df[eq_df['품목명'].apply(lambda x: not self._is_consumable_material(str(x).strip(), ''))]
                
                unique_items = equipment_df[['품목명', 'SN', '모델명']].drop_duplicates()
                for _, row in unique_items.iterrows():
                    name = str(row.get('품목명', '')).strip()
                    sn = str(row.get('SN', '')).strip()
                    model = str(row.get('모델명', '')).strip()
                    
                    if name.lower() == 'nan': name = ''
                    if sn.lower() == 'nan': sn = ''
                    if model.lower() == 'nan': model = ''
                    
                    if name or sn or model:
                        if not query or query in name.lower() or query in sn.lower() or query in model.lower():
                            tree.insert('', tk.END, values=(name, sn, model))

        def on_select(e=None):
            selection = tree.selection()
            if not selection: return
            
            vals = tree.item(selection[0])['values']
            name, sn, model = str(vals[0]), str(vals[1]), str(vals[2])
            
            # Format: [Model] Name (SN)
            display_parts = []
            if model and model != 'None' and model != '': display_parts.append(model)
            if name and name != model and name != 'None' and name != '': display_parts.append(name)
            if sn and sn != 'None' and sn != '': display_parts.append(f"({sn})")
            
            display_name = " ".join(display_parts).strip()
            
            if display_name:
                self.cb_daily_equip.delete(0, tk.END)
                self.cb_daily_equip.insert(0, display_name)
                dlg.destroy()

        def on_direct_entry():
            query = search_var.get().strip()
            if not query:
                messagebox.showwarning("입력 오류", "등록할 장비명을 입력해주세요.")
                return
            
            # Use the query as the equipment name directly
            if query not in self.equipments:
                self.equipments.append(query)
                self.refresh_ui_for_list_change('equipments')
            dlg.destroy()

        search_var.trace_add("write", refresh_list)
        tree.bind('<Double-1>', on_select)
        
        refresh_list()

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="선택 완료", command=on_select, width=15).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="입력 명칭으로 신규 등록", command=on_direct_entry, width=22).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="취소", command=dlg.destroy, width=10).pack(side='left', padx=5)

    def open_material_search_dialog(self, target_form='transaction'):
        """Unified dialog to search materials, edit them, or select for forms"""
        dlg = tk.Toplevel(self.root)
        dlg.title("자재 마스터 검색 및 관리")
        dlg.geometry("900x700")
        dlg.transient(self.root)
        dlg.grab_set()

        main_frame = ttk.Frame(dlg, padding=20)
        main_frame.pack(fill='both', expand=True)

        header_text = "전체 재고 자재 검색 및 수정" if target_form == 'registration' else "자재 검색 및 선택"
        ttk.Label(main_frame, text=header_text, font=('Malgun Gothic', 12, 'bold')).pack(pady=(0, 10))
        
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill='x', pady=5)
        ttk.Label(search_frame, text="🔍 검색어:").pack(side='left', padx=5)
        search_var = tk.StringVar()
        search_ent = ttk.Entry(search_frame, textvariable=search_var)
        search_ent.pack(side='left', fill='x', expand=True, padx=5)
        search_ent.focus_set()

        # Treeview for results
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill='both', expand=True)
        
        columns = ('id', 'name', 'model', 'sn', 'spec', 'unit', 'stock')
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        tree.heading('id', text='ID')
        tree.heading('name', text='품목명')
        tree.heading('model', text='모델명')
        tree.heading('sn', text='SN')
        tree.heading('spec', text='규격')
        tree.heading('unit', text='단위')
        tree.heading('stock', text='재고')
        
        tree.column('id', width=50)
        tree.column('name', width=180)
        tree.column('model', width=150)
        tree.column('sn', width=100)
        tree.column('spec', width=100)
        tree.column('unit', width=60)
        tree.column('stock', width=60)
        
        sb_y = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb_y.set)
        
        tree.pack(side='left', fill='both', expand=True)
        sb_y.pack(side='right', fill='y')

        def refresh_list(*args):
            query = search_var.get().strip().lower()
            for item in tree.get_children():
                tree.delete(item)
            
            if not self.materials_df.empty:
                for idx, row in self.materials_df.iterrows():
                    # [STRICT] Filter out inactive items (Active=0) from search dialog
                    # Use a robust check that handles both Series and scalars
                    active_val = row.get('Active', 1)
                    num_active = pd.to_numeric(active_val, errors='coerce')
                    is_active = (num_active > 0.5) if pd.notna(num_active) else True
                    
                    if not is_active:
                        continue
                    
                    name = str(row.get('품목명', '')).strip()
                    sn = str(row.get('SN', '')).strip()
                    model = str(row.get('모델명', '')).strip()
                    spec = str(row.get('규격', '')).strip()
                    unit = str(row.get('관리단위', '')).strip()
                    stock = str(row.get('수량', '0'))
                    # [NEW] Format MaterialID as integer for display
                    mid_raw = row.get('MaterialID', '')
                    mid = ""
                    try:
                        if pd.notna(mid_raw) and str(mid_raw).strip():
                            mid = str(int(float(str(mid_raw).strip())))
                    except:
                        mid = str(mid_raw)
                    
                    if spec == "자동등록": continue
                    
                    # Clean nans
                    name = '' if name.lower() == 'nan' else name
                    sn = '' if sn.lower() == 'nan' else sn
                    model = '' if model.lower() == 'nan' else model
                    spec = '' if spec.lower() == 'nan' else spec
                    
                    if name or sn or model:
                        display_name = name if is_active else f"[비활성] {name}"
                        row_str = f"{name} {sn} {model} {spec} {unit}".lower()
                        if not query or query in row_str:
                            tag = 'active' if is_active else 'inactive'
                            tree.insert('', tk.END, values=(mid, display_name, model, sn, spec, unit, stock), tags=(tag, idx))
                
                tree.tag_configure('inactive', foreground='gray')
                tree.tag_configure('active', foreground='black')

        def on_select(e=None):
            selection = tree.selection()
            if not selection: return
            
            vals = tree.item(selection[0])['values']
            name = str(vals[1])
            model = str(vals[2])
            sn = str(vals[3])
            
            if target_form == 'transaction':
                self.cb_material.set(name)
                self.on_material_selected()
                dlg.destroy()
            elif target_form == 'daily_usage':
                self.cb_daily_material.delete(0, tk.END)
                self.cb_daily_material.insert(0, name)
                dlg.destroy()
            else:
                # Registration form
                self.cb_item_name.set(name)
                if hasattr(self, 'cb_model'): self.cb_model.set(model)
                if hasattr(self, 'ent_sn'): 
                    self.ent_sn.delete(0, tk.END)
                    self.ent_sn.insert(0, sn)
                dlg.destroy()

        def edit_material():
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("선택 오류", "수정할 항목을 선택해주세요.")
                return
            
            item_id = tree.item(selection[0])['values'][0]
            # Find in DF
            mask = self.materials_df['MaterialID'].astype(str) == str(item_id)
            if not mask.any(): return
            
            row_idx = self.materials_df.index[mask][0]
            mat_data = self.materials_df.loc[row_idx]

            edit_win = tk.Toplevel(dlg)
            edit_win.title(f"자재 정보 수정 - ID: {item_id}")
            edit_win.geometry("400x500")
            edit_win.transient(dlg)
            edit_win.grab_set()

            f = ttk.Frame(edit_win, padding=20)
            f.pack(fill='both', expand=True)

            fields = [
                ("품목명", "품목명"), ("모델명", "모델명"), ("SN", "SN"), 
                ("규격", "규격"), ("단위", "관리단위"), ("수량", "수량"), ("재고하한", "재고하한")
            ]
            entries = {}
            for i, (label, col) in enumerate(fields):
                ttk.Label(f, text=f"{label}:").grid(row=i, column=0, padx=5, pady=5, sticky='w')
                ent = ttk.Entry(f, width=30)
                ent.grid(row=i, column=1, padx=5, pady=5, sticky='ew')
                val = str(mat_data.get(col, ''))
                if val.lower() == 'nan': val = ''
                ent.insert(0, val)
                entries[col] = ent

            def save_edit():
                for col, ent in entries.items():
                    val = ent.get().strip()
                    if col in ['수량', '재고하한']:
                        try: val = float(val) if val else 0.0
                        except: val = 0.0
                    self.materials_df.at[row_idx, col] = val
                
                self.save_data()
                self.update_stock_view()
                self.update_material_combo()
                refresh_list()
                messagebox.showinfo("수정 완료", "자재 정보가 업데이트되었습니다.")
                edit_win.destroy()

            ttk.Button(edit_win, text="변경사항 저장", command=save_edit).pack(pady=10)

        def add_new_material():
            """Small dialog to add a missing material quickly"""
            add_win = tk.Toplevel(dlg)
            add_win.title("자재 신규 등록")
            add_win.geometry("450x350")
            add_win.transient(dlg)
            add_win.grab_set()

            form_f = ttk.Frame(add_win, padding=20)
            form_f.pack(fill='both', expand=True)

            entries = {}
            labels = [("품목명", "name"), ("모델명", "model"), ("SN", "sn"), ("규격", "spec"), ("단위", "unit")]
            
            for i, (label, key) in enumerate(labels):
                ttk.Label(form_f, text=f"{label}:").grid(row=i, column=0, padx=5, pady=5, sticky='w')
                ent = ttk.Entry(form_f, width=40)
                ent.grid(row=i, column=1, padx=5, pady=5, sticky='ew')
                entries[key] = ent
                if key == "name":
                    ent.insert(0, search_var.get().strip())
                    ent.focus_set()
                if key == "unit":
                    ent.insert(0, "매")

            def save_new():
                data = {k: e.get().strip() for k, e in entries.items()}
                if not data['name']:
                    messagebox.showwarning("오류", "품목명은 필수입니다.")
                    return
                
                new_id = 10001
                if not self.materials_df.empty:
                    valid_ids = pd.to_numeric(self.materials_df['MaterialID'], errors='coerce').dropna()
                    if not valid_ids.empty:
                        new_id = int(valid_ids.max()) + 1

                new_row = {
                    'MaterialID': new_id,
                    '품목명': data['name'],
                    '모델명': data['model'],
                    'SN': data['sn'],
                    '규격': data['spec'],
                    '관리단위': data['unit'],
                    'Active': 1,
                    '수량': 0,
                    '재고하한': 0
                }
                for col in self.materials_df.columns:
                    if col not in new_row: new_row[col] = ''
                
                self.materials_df = pd.concat([self.materials_df, pd.DataFrame([new_row])], ignore_index=True)
                self.save_data()
                self.update_material_combo()
                self.update_stock_view()
                refresh_list()
                messagebox.showinfo("완료", f"'{data['name']}' 자재가 성공적으로 등록되었습니다.")
                add_win.destroy()

            ttk.Button(add_win, text="등록 완료", command=save_new).pack(pady=10)

        search_var.trace_add("write", refresh_list)
        tree.bind('<Double-1>', on_select)
        refresh_list()

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="✅ 선택 완료", command=on_select, width=15).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="✏️ 정보 수정", command=edit_material, width=15).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="➕ 직접 등록", command=add_new_material, width=15).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="취소", command=dlg.destroy, width=10).pack(side='left', padx=5)

    def open_preferred_item_search_dialog(self):
        """Deprecated in favor of unified open_material_search_dialog"""
        self.open_material_search_dialog(target_form='registration')

    def _get_merged_memo_and_note(self):
        """[NEW] Collects user-entered note and all open memo contents for archiving"""
        main_note = ""
        if hasattr(self, 'ent_daily_note'):
            main_note = self.ent_daily_note.get().strip()
            
        memo_archive = []
        if hasattr(self, 'memos'):
            for key, m in self.memos.items():
                try:
                    title = m['title_entry'].get().strip()
                    content = m['text_widget'].get('1.0', 'end-1c').strip()
                    if content:
                        memo_archive.append(content)
                except: continue
        
        if not memo_archive:
            return main_note
            
        archive_text = " | ".join(memo_archive)
        return f"{main_note} | {archive_text}" if main_note else archive_text

    def load_daily_usage_to_form(self, record):
        """[NEW] Load a database record back into the input form for editing or report generation"""
        try:
            def to_f_local(v):
                try: return float(str(v).replace(',', '')) if v else 0.0
                except: return 0.0

            # 1. Basic Info
            if 'Date' in record:
                try: self.ent_daily_date.set_date(pd.to_datetime(record['Date']))
                except: pass
            if '업체명' in record:
                comp_val = self.clean_nan(record['업체명']).split(' [')[0]
                self.cb_daily_company.set(comp_val)
            if 'Site' in record:
                self.cb_daily_site.set(self.clean_nan(record['Site']))
            if '구분' in record:
                self.ndt_loc_type_var.set(self.clean_nan(record['구분']))
            if '적용코드' in record:
                self.ent_daily_applied_code.delete(0, tk.END)
                self.ent_daily_applied_code.insert(0, self.clean_nan(record['적용코드']))
            if '장비명' in record:
                self.cb_daily_equip.delete(0, tk.END)
                self.cb_daily_equip.insert(0, self.clean_nan(record['장비명']))
            if '성적서번호' in record:
                self.ent_daily_report_no.delete(0, tk.END)
                self.ent_daily_report_no.insert(0, self.clean_nan(record['성적서번호']))
            if '검사품명' in record:
                self.ent_daily_inspection_item.delete(0, tk.END)
                self.ent_daily_inspection_item.insert(0, self.clean_nan(record['검사품명']))
            if '검사방법' in record:
                method = self.clean_nan(record.get('검사방법', ''))
                self.cb_daily_test_method.set(method)
                
                # [FIX] Explicitly show/hide NDT related frames to ensure they appear
                if method in ["MT", "PT"]:
                    if hasattr(self, 'ndt_frame'):
                        try: self.ndt_frame.grid()
                        except: pass
                else:
                    if hasattr(self, 'ndt_frame'):
                        try: self.ndt_frame.grid_remove()
                        except: pass
                
                if method in ["RT", "UT", "PT", "PAUT"]:
                    if hasattr(self, 'ndt_calc_frame'):
                        try:
                            self.ndt_calc_frame.grid(row=9, column=0, columnspan=4, sticky='ew', pady=(5,0))
                            self.ndt_calc_frame.lift()
                        except: pass
                    if method == "RT":
                        if hasattr(self, 'rtk_grid'):
                            try: self.rtk_grid.grid()
                            except: pass
                    else:
                        if hasattr(self, 'rtk_grid'):
                            try: self.rtk_grid.grid_remove()
                            except: pass
                else:
                    if hasattr(self, 'ndt_calc_frame'):
                        try: self.ndt_calc_frame.grid_remove()
                        except: pass
                    if hasattr(self, 'rtk_grid'):
                        try: self.rtk_grid.grid_remove()
                        except: pass
                        
                self.cb_daily_test_method.event_generate('<<ComboboxSelected>>')
            if 'Unit' in record:
                self.cb_daily_unit.set(self.clean_nan(record['Unit']))
            elif '단위' in record:
                self.cb_daily_unit.set(self.clean_nan(record['단위']))
            
            # 2. Quantities & Costs
            def set_val(ent, key):
                val = record.get(key, 0)
                ent.delete(0, tk.END)
                if isinstance(val, (int, float)):
                    if val != 0: ent.insert(0, f"{val:,.0f}" if any(x in key for x in ['비', '가', '일식']) else str(val))
                else: 
                    ent.insert(0, str(val))

            set_val(self.ent_daily_test_amount, '검사량')
            set_val(self.ent_daily_unit_price, '단가')
            set_val(self.ent_daily_travel_cost, '출장비')
            set_val(self.ent_daily_test_fee, '검사비')
            set_val(self.ent_daily_meal_cost, '일식')
            if '작업형태' in record and record['작업형태']: self.ndt_work_time_var.set(record['작업형태'])
            if '조건1' in record and record['조건1']: 
                if record.get('검사방법') == 'RT': self.ndt_source_var.set(record['조건1'])
                else: self.ndt_pipe_var.set(record['조건1'])
            if '조건2' in record and record['조건2']: self.ndt_thickness_var.set(record['조건2'])
            def get_valid_rate(key, default_val):
                if key in record:
                    val = record[key]
                    if pd.notna(val) and str(val).strip() and str(val).lower() != 'nan':
                        try: return float(val)
                        except: pass
                return default_val
            
            self.ndt_overhead_var.set(get_valid_rate('제경비율', 110.0))
            self.ndt_tech_var.set(get_valid_rate('기술료율', 20.0))

            # NDT 상세 조건 필드 로딩
            if '관경(Inch)' in record and pd.notna(record['관경(Inch)']):
                self.ndt_report_pipe_var.set(str(record['관경(Inch)']).replace('.0', '') if str(record['관경(Inch)']).endswith('.0') else str(record['관경(Inch)']))
            else:
                self.ndt_report_pipe_var.set("")
                
            insp_type = self.clean_nan(record.get('검사구분', 'ORI')).upper()
            if not insp_type: insp_type = 'ORI'
            if insp_type == 'REP':
                self.ndt_rep_joint_var.set(self.clean_nan(record.get('조인트수', '')))
                self.ndt_rep_qty_var.set(self.clean_nan(record.get('Usage', '')))
            else:
                self.ndt_ori_joint_var.set(self.clean_nan(record.get('조인트수', '')))
                self.ndt_ori_qty_var.set(self.clean_nan(record.get('Usage', '')))
                
            if '불량수' in record and pd.notna(record['불량수']):
                self.ndt_rej_joint_var.set(str(record['불량수']))
            else:
                self.ndt_rej_joint_var.set("")
            
            # 3. Material
            if 'MaterialID' in record:
                mat_id_val = record['MaterialID']
                disp_name = self.get_material_display_name(mat_id_val)
                # [FIX] If display name is empty but MaterialID is a non-NaN string (e.g. 'JIREH Scanner'),
                # show the raw MaterialID string as a fallback
                if not disp_name:
                    try:
                        if not pd.isna(mat_id_val) and str(mat_id_val).strip().lower() not in ('nan', ''):
                            disp_name = str(mat_id_val).strip()
                    except: pass
                print(f"[DEBUG] load_to_form: MaterialID={repr(mat_id_val)}, disp={repr(disp_name)}")
                if isinstance(self.cb_daily_material, ttk.Combobox):
                    self.cb_daily_material.set(disp_name)
                else:
                    self.cb_daily_material.delete(0, tk.END)
                    self.cb_daily_material.insert(0, disp_name)

            # 4. Workers
            # First, clear existing workers if needed or just overwrite
            for i in range(1, 11):
                group = getattr(self, f'worker_group{i}', None)
                if not group: continue
                u_key = 'User' if i == 1 else f'User{i}'
                wt_key = 'WorkTime' if i == 1 else f'WorkTime{i}'
                ot_key = 'OT' if i == 1 else f'OT{i}'
                m_key = 'Meal' if i == 1 else f'Meal{i}'
                
                group.cb_name.set(self.clean_nan(record.get(u_key, '')))
                # [FIX] Use set_time to properly parse (Shift) Time string and populate both shift and time widgets
                group.set_time(self.clean_nan(record.get(wt_key, '')))
                
                group.ent_ot.delete(0, tk.END)
                group.ent_ot.insert(0, self.clean_nan(record.get(ot_key, '')))
                
                group.ent_meal.delete(0, tk.END)
                group.ent_meal.insert(0, self.clean_nan(record.get(m_key, '')))

            # 5. RTK

            for cat, ent in self.rtk_entries.items():
                if cat == "총계": continue
                val = record.get(f'RTK_{cat}', 0)
                ent.delete(0, tk.END)
                if abs(to_f_local(val)) > 0.001: ent.insert(0, str(val))
            self.calculate_rtk_total()


            # 6. Vehicle
            v_no_raw = self.clean_nan(record.get('차량번호', ''))
            v_insp_raw = self.clean_nan(record.get('차량점검', ''))
            v_mileage_raw = self.clean_nan(record.get('주행거리', ''))
            v_remarks_raw = self.clean_nan(record.get('차량비고', ''))

            v_no_list = [x.strip() for x in str(v_no_raw).split('||')] if v_no_raw else []
            v_insp_list = [x.strip() for x in str(v_insp_raw).split('||')] if v_insp_raw else []
            v_mileage_list = [x.strip() for x in str(v_mileage_raw).split('||')] if v_mileage_raw else []
            v_remarks_list = [x.strip() for x in str(v_remarks_raw).split('||')] if v_remarks_raw else []
            
            max_v_count = max(len(v_no_list), len(v_insp_list), len(v_mileage_list), len(v_remarks_list))
            
            # 기존 플로팅 창 닫기 및 고정 패널 비우기
            if hasattr(self, 'draggable_items'):
                for key, cont in list(self.draggable_items.items()):
                    w = getattr(cont, '_widget', None)
                    if not w: continue
                    if type(w).__name__ == 'VehicleInspectionWidget':
                        self.remove_box(key)
            if hasattr(self, 'fixed_vehicle_widget'):
                self.fixed_vehicle_widget.reset_fields()

            for i in range(max_v_count):
                cur_no = v_no_list[i] if i < len(v_no_list) else ""
                cur_insp = v_insp_list[i] if i < len(v_insp_list) else ""
                cur_mileage = v_mileage_list[i] if i < len(v_mileage_list) else ""
                cur_remarks = v_remarks_list[i] if i < len(v_remarks_list) else ""
                
                if not (cur_no or cur_insp or cur_mileage or cur_remarks):
                    continue

                v_parsed = {'vehicle_info': cur_no, 'mileage': cur_mileage, 'remarks': cur_remarks}
                if ':' in cur_insp:
                    for pair in cur_insp.split('|'):
                        if ':' in pair:
                            k, v_val = pair.split(':', 1)
                            v_parsed[k.strip()] = v_val.strip()
                elif cur_insp and cur_insp != 'nan':
                    for k in cur_insp.split(','):
                        k_clean = k.strip()
                        if k_clean:
                            if 'locking' in k_clean: v_parsed[k_clean] = '잠금'
                            elif 'cleaning' in k_clean: v_parsed[k_clean] = '함'
                            else: v_parsed[k_clean] = '양호'
                
                if i == 0 and hasattr(self, 'fixed_vehicle_widget'):
                    v_widget = self.fixed_vehicle_widget
                    v_widget.set_data(v_parsed)
                else:
                    if hasattr(self, 'add_vehicle_inspection_box'):
                        v_widget = self.add_vehicle_inspection_box()
                        if hasattr(v_widget, 'set_data'):
                            v_widget.set_data(v_parsed)
            
            # 7. NDT Chemicals
            if hasattr(self, 'ndt_company_entries') and self.ndt_company_entries:
                # [STABILITY] Clear first if multiple companies, but for now focus on the primary/first entry
                first = self.ndt_company_entries[0]
                ndt_names = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
                for name in ndt_names:
                    col_name = f"NDT_{name}"
                    if col_name in record:
                        val = record[col_name]
                        if name in first:
                            first[name].delete(0, tk.END)
                            if abs(to_f_local(val)) > 0.001:
                                first[name].insert(0, str(val))
                
                # Also load company if available
                if '회사코드' in record:
                    first['_company'].set(self.clean_nan(record.get('회사코드')))

            # 8. Note
            note_val = self.clean_nan(record.get('Note', ''))
            if 'Note' in record:
                self.ent_daily_note.delete(0, tk.END)
                # [REMOVED] 사용자 요청으로 과거 비고(Note) 내용을 불러오지 않음
                # self.ent_daily_note.insert(0, note_val)
                
            # 9. DailyMemo
            # 사용자가 저장한 메모를 불러오도록 복구 (과거 기록 호환성 포함)
            memo_val = self.clean_nan(record.get('DailyMemo', ''))
            if not memo_val and note_val:
                if ' | ' in str(note_val):
                    parts = str(note_val).split(' | ', 1)
                    if len(parts) > 1:
                        memo_val = parts[1]
                else:
                    memo_val = note_val
                    
            if hasattr(self, 'main_memo_text') and self.main_memo_text:
                try:
                    self.main_memo_text.delete('1.0', tk.END)
                    if memo_val:
                        self.main_memo_text.insert(tk.END, memo_val)
                except: pass

        except Exception as e:
            print(f"Error loading record to form: {e}")

    def clear_daily_usage_form_all(self, keep_date=False):
        """현장별 일일 사용량 입력 폼의 모든 필드를 초기화합니다."""
        # 1. 콤보박스 선택 해제
        for cb in [self.cb_daily_company, self.cb_daily_site, 
                   self.cb_daily_test_method, self.cb_daily_unit]:
            cb.set('')
        
        # 2. 날짜 초기화 (연속 기입 시 편의를 위해 keep_date 옵션 제공)
        if not keep_date:
            try:
                self.ent_daily_date.set_date(datetime.date.today())
            except: pass
        
        # 3. 모든 Entry 필드 비우기
        for ent in [self.ent_daily_inspection_item, self.ent_daily_applied_code, 
                   self.ent_daily_equip, self.cb_daily_material, # Both are now Entries
                   self.ent_daily_test_amount, self.ent_daily_unit_price, 
                   self.ent_daily_report_no, self.ent_daily_travel_cost, 
                   self.ent_daily_meal_cost, self.ent_daily_test_fee, self.ent_daily_note]:
            ent.delete(0, tk.END)
            
        # 기본값 복구
        self.cb_daily_unit.set('매')
        self.ndt_work_time_var.set("일반")
        self.ndt_source_var.set("Se-75 (1.0)")
        self.ndt_thickness_var.set("조건없음 (1.0)")
        self.ndt_pipe_var.set("250mm 초과 [10인치 이상] (1.0)")
        self.ndt_overhead_var.set(110.0)
        self.ndt_tech_var.set(20.0)
        self.ndt_ori_joint_var.set("")
        self.ndt_ori_qty_var.set("")
        self.ndt_rep_joint_var.set("")
        self.ndt_rep_qty_var.set("")
        self.ndt_rej_joint_var.set("")
        self.ndt_report_pipe_var.set("")
        self.ndt_calc_frame.grid_remove()
        self.rtk_grid.grid_remove() # [NEW] Hide RTK on clear
        self.ndt_frame.grid_remove() # [NEW] Hide NDT frame on clear
        
        # 4. NDT 자재 및 RTK 필드 초기화
        if hasattr(self, 'ndt_company_entries'):
            for entries in self.ndt_company_entries:
                for mat in ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]:
                    if mat in entries: entries[mat].delete(0, tk.END)
        
        for ent in self.rtk_entries.values():
            if ent.winfo_exists():
                ent.config(state='normal')
                ent.delete(0, tk.END)
                if ent == self.rtk_entries.get("총계"): 
                    ent.config(state='readonly')
        
        # 5. 차량 점검 섹션 초기화 (중복 방지 및 고정 패널 사용)
        if hasattr(self, 'draggable_items'):
            for key, cont in list(self.draggable_items.items()):
                w = getattr(cont, '_widget', None)
                if not w: continue
                lbl_text = ""
                if hasattr(cont, '_label_widget'):
                    lbl_text = cont._label_widget.cget('text')
                is_vehicle = (type(w).__name__ == 'VehicleInspectionWidget') or ('차량' in lbl_text and '점검' in lbl_text) or ('상시' in lbl_text and '차량' in lbl_text)
                if is_vehicle:
                    print(f"[DEBUG] Removing legacy floating vehicle box: {lbl_text}")
                    self.remove_box(key)

        if hasattr(self, 'fixed_vehicle_widget'):
            if hasattr(self.fixed_vehicle_widget, 'reset_fields'):
                self.fixed_vehicle_widget.reset_fields()
                
        # 6. [REMOVED] 메모 패널 초기화 방지 (사용자 요청)
                    
        # 6. 작업자 섹션 초기화
        for i in range(1, 11):
            group = getattr(self, f'worker_group{i}', None)
            if group:
                group.cb_name.set('')
                group.cb_shift.set('주간')
                group.ent_worktime.set('')
                group.ent_ot.delete(0, tk.END)

    def add_daily_usage_entry(self):
        """현장별 일일 사용량 단일 등록 (리팩토링 버전)"""
        try:
            # 1. 기본 정보 및 유효성 검사
            date_val = self.ent_daily_date.get_date()
            site = self.cb_daily_site.get().strip()
            if not site:
                messagebox.showwarning("입력 오류", "현장명을 입력해주세요.")
                return

            # [NEW VALIDATION START]
            # 1-1. 차량 운행 유효성 검사
            if hasattr(self, 'vehicle_widget'):
                veh_data = self.vehicle_widget.get_data()
                if veh_data.get('vehicle_info'):
                    if not veh_data.get('mileage'):
                        if not messagebox.askyesno("입력 확인", "차량 정보는 선택되었으나 주행거리가 입력되지 않았습니다.\n이대로 저장하시겠습니까?"):
                            return

            # 1-2. 투입 인원 대비 검사 물량 검사
            has_workers = False
            if hasattr(self, 'worker_groups'):
                for wg in self.worker_groups:
                    if wg.get_worker() and wg.get_time():
                        has_workers = True
                        break
            
            has_ndt_qty = False
            if hasattr(self, 'ndt_company_entries'):
                for comp_dict in self.ndt_company_entries:
                    for mat, entry_widget in comp_dict.items():
                        if mat == '_company': continue
                        try:
                            qty = float(entry_widget.get().strip() or 0)
                            if qty > 0:
                                has_ndt_qty = True
                                break
                        except: pass
                    if has_ndt_qty: break

            if has_workers and not has_ndt_qty:
                if not messagebox.askyesno("입력 확인", "투입된 인원(작업시간)은 있으나 검사 물량(NDT 수량)이 모두 0입니다.\n결측치일 가능성이 있습니다. 이대로 저장하시겠습니까?"):
                    return
            # [NEW VALIDATION END]

            # 2. 품목 ID 확인 (디스플레이 명칭 기반 매칭)
            mat_display = self.cb_daily_material.get().strip()
            mat_id = ""
            if mat_display:
                for _, row in self.materials_df.iterrows():
                    if self.get_material_display_name(row['MaterialID']) == mat_display:
                        mat_id = row['MaterialID']
                        # 휴면 계정 자동 활성화
                        if row.get('Active', 1) == 0:
                            self.materials_df.loc[self.materials_df['MaterialID'] == mat_id, 'Active'] = 1
                        break
                
                # [FIX] Only auto-register if it's a CONSUMABLE (Drug/Film). Equipment should NOT be registered.
                if not mat_id and mat_display:
                    if self._is_consumable_material(mat_display, self.cb_daily_test_method.get().strip()):
                        # [UX IMPROVEMENT] 등록되지 않은 소모품 감지 시, 사용자의 승인을 받도록 알림 팝업 추가
                        if messagebox.askyesno("신규 자재 등록", f"등록되지 않은 소모품 자재명('{mat_display}')입니다.\n재고 마스터에 신규 자재로 등록하시겠습니까?\n\n(아니오를 누르면 작업일보에는 저장되지만 재고 연동(차감) 대상에서 제외됩니다.)"):
                            mat_id = self.register_new_material(mat_display, warehouse='현장', 규격='자동등록')
                        else:
                            mat_id = mat_display
                    else:
                        # For equipment or non-consumables, just store the name as ID (skips stock tracking)
                        mat_id = mat_display

            # 3. 중복 저장 방지 체크
            if not self.daily_usage_df.empty:
                try:
                    # Normalize date for comparison
                    check_date = pd.to_datetime(date_val).date()
                    # Ensure MaterialID is matched correctly (exact string match)
                    existing = self.daily_usage_df[
                        (pd.to_datetime(self.daily_usage_df['Date']).dt.date == check_date) & 
                        (self.daily_usage_df['Site'] == site) & 
                        (self.daily_usage_df['MaterialID'] == mat_id) &
                        (self.daily_usage_df['검사품명'].astype(str).str.strip() == self.ent_daily_inspection_item.get().strip()) &
                        (self.daily_usage_df['적용코드'].astype(str).str.strip() == self.ent_daily_applied_code.get().strip())
                    ]
                    if not existing.empty:
                        answer = messagebox.askyesnocancel("중복 확인", 
                            f"이미 {date_val} 날짜에 '{site}' 현장의 '{mat_display}' 품목 기록이 {len(existing)}건 존재합니다.\n\n"
                            "기존 기록을 덮어쓰시겠습니까?\n\n"
                            "'예(Y)': 기존 기록 덮어쓰기 (수정)\n"
                            "'아니오(N)': 새로운 기록으로 추가 저장\n"
                            "'취소(Cancel)': 작업 취소")
                        
                        if answer is None:
                            return
                        elif answer is True:
                            # 덮어쓰기 위해 기존 기록 삭제 및 트랜잭션 롤백
                            for idx in existing.index:
                                entry = self.daily_usage_df.loc[idx]
                                site_name = entry.get('Site', '')
                                usage_date = pd.to_datetime(entry.get('Date'))
                                mat_id_to_delete = entry.get('MaterialID', '')
                                if not self.transactions_df.empty:
                                    trans_mask = (
                                        (pd.to_datetime(self.transactions_df['Date'], errors='coerce').dt.normalize() == pd.to_datetime(usage_date).normalize()) &
                                        (self.transactions_df['Site'].astype(str) == str(site_name)) &
                                        (self.transactions_df['MaterialID'].astype(str) == str(mat_id_to_delete)) &
                                        (self.transactions_df['Type'] == 'OUT') &
                                        (self.transactions_df['Note'].str.contains(f"{site_name} 현장 사용", na=False, regex=False))
                                    )
                                    self.transactions_df = self.transactions_df[~trans_mask]
                            
                            self.daily_usage_df = self.daily_usage_df.drop(existing.index)
                            self.daily_usage_df = self.daily_usage_df.reset_index(drop=True)
                except Exception as e:
                    print(f"DEBUG: Duplicate check failed: {e}")

            # [DEBUG] Trace what is being saved
            print(f"[DEBUG SAVE] mat_display={repr(mat_display)}, mat_id={repr(mat_id)}")
            print(f"[DEBUG SAVE] cb_daily_material.get()={repr(self.cb_daily_material.get())}")
            print(f"[DEBUG SAVE] cb_daily_equip.get()={repr(self.cb_daily_equip.get())}")
            if hasattr(self, 'vehicle_boxes') and self.vehicle_boxes:
                vd = self.vehicle_boxes[0].get_data()
                print(f"[DEBUG SAVE] vehicle[0]: no={repr(vd.get('vehicle_info',''))}, km={repr(vd.get('mileage',''))}")
            else:
                print(f"[DEBUG SAVE] vehicle_boxes: none/empty")

            # 4. 핵심 로직 실행 (단건 저장)
            saved_count = self._add_single_usage_record_logic(mat_id, date_val, site, auto_save=True)
            if saved_count > 0:
                # [SMART VISIBILITY] Adjust filters to ensure the NEW record is visible
                try:
                    current_start = self.ent_daily_start_date.get_date()
                    current_end = self.ent_daily_end_date.get_date()
                    
                    if date_val < current_start:
                        self.ent_daily_start_date.set_date(date_val)
                    if date_val > current_end:
                        self.ent_daily_end_date.set_date(date_val)
                except: pass
                
                # 성공 시 필드 초기화 및 뷰 갱신
                if hasattr(self, 'ndt_company_entries'):
                    for company_entries in self.ndt_company_entries:
                        for name in self.ndt_materials_all:
                            if name in company_entries:
                                company_entries[name].delete(0, tk.END)
                
                # RTK 필드 초기화
                for cat, ent in self.rtk_entries.items():
                    if cat != "총계":
                        ent.delete(0, tk.END)
                self.rtk_entries["총계"].config(state='normal')
                self.rtk_entries["총계"].delete(0, tk.END)
                self.rtk_entries["총계"].config(state='readonly')
                
                # [V13_RESET_REQUESTED] Clearing fields after successful entry based on user request
                # [FIX] Keep the current date for consecutive entries on the same day
                self.clear_daily_usage_form_all(keep_date=True)

                # Reset focus to first logical field (Company)
                self.cb_daily_company.focus_set()

                # [V13.1] Keep workers too for consecutive entries
                # for i in range(1, 11):
                #     group = getattr(self, f'worker_group{i}', None)
                #     if group:
                #         group.cb_name.set('')
                #         group.cb_shift.set('')
                #         group.ent_worktime.set('')
                #         group.ent_ot.delete(0, tk.END)

                # NDT 섹션 초기화
                if hasattr(self, 'ndt_company_entries'):
                    while len(self.ndt_company_entries) > 1:
                        self.remove_last_ndt_company()
                    if self.ndt_company_entries:
                        first = self.ndt_company_entries[0]
                        first['_company'].set('')
                        for k in self.ndt_materials_all:
                            if k in first: first[k].delete(0, tk.END)

                # 차량 점검 필드 초기화 (safeguarded)
                # [NEW] 현장별 저장 시 차량, 목록, 메모 레이블창 자동 닫기 (User request)
                for key in list(getattr(self, 'vehicle_inspections', {}).keys()):
                    self.remove_box(key)
                for key in list(getattr(self, 'checklists', {}).keys()):
                    self.remove_box(key)
                for key in list(getattr(self, 'memos', {}).keys()):
                    self.remove_box(key)

                self.update_daily_usage_view()
                self.update_recent_entries_view()
                self.refresh_inquiry_filters()
                self.update_stock_view()
                
                # [NEW] Scroll to the bottom to see new entries
                if hasattr(self, 'daily_usage_tree'):
                    self.daily_usage_tree.yview_moveto(1.0)
                
                messagebox.showinfo("성공", f"{saved_count}건의 기록이 안전하게 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"기록 저장 중 오류 발생: {e}")

    def load_report_mapping(self):
        """Load report mapping from JSON file or return defaults"""
        mapping_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'resources', 'report_mapping.json')
        if not os.path.exists(mapping_path):
            # Return defaults matching the manager's default_mapping
            return {
                'header': {'date': 'F2'},
                'general': {
                    'company': 'E5', 'project_name': 'E6', 'standard': 'E7', 'equipment': 'E8',
                    'report_no': 'K5', 'inspection_item': 'K6', 'inspector': 'K7', 'car_no': 'K8'
                },
                'methods': {
                    'RT': {'row': '12'}, 'UT': {'row': '13'}, 'MT': {'row': '14'}, 'PT': {'row': '15'},
                    'HT': {'row': '16'}, 'VT': {'row': '17'}, 'LT': {'row': '18'}, 'ET': {'row': '19'},
                    'PAUT': {'row': '20'}
                },
                'rtk': {
                    'center_miss': 'C32', 'density': 'E32', 'marking_miss': 'G32', 'film_mark': 'I32',
                    'handling': 'K32', 'customer_complaint': 'M32', 'etc': 'O32', 'total': 'Q32'
                },
                'ot': {
                    'row1_name': 'B38', 'row1_hours': 'K38',
                    'row2_name': 'B39', 'row2_hours': 'K39'
                },
                'materials': {
                    'RT T200': '43', 'RT AA400': '44', 'MT WHITE': '46', 'MT 7C-BLACK': '47',
                    'PT Penetrant': '48', 'PT Cleaner': '49', 'PT Developer': '50'
                }
            }
        try:
            with open(mapping_path, 'r', encoding='utf-8') as f:
                import json
                mapping = json.load(f)
                
                # Robust cleaning of method rows (e.g. 'E13' -> 13)
                if 'methods' in mapping:
                    import re
                    for m_key, m_val in mapping['methods'].items():
                        if isinstance(m_val, dict) and 'row' in m_val:
                            row_str = str(m_val['row'])
                            match = re.search(r'\d+', row_str)
                            if match: m_val['row'] = int(match.group())
                return mapping
        except:
            return {}

    def save_report_mapping(self, mapping):
        """Save report mapping to JSON file"""
        resources_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'resources')
        if not os.path.exists(resources_dir): 
            try: os.makedirs(resources_dir)
            except: pass
        mapping_path = os.path.join(resources_dir, 'report_mapping.json')
        try:
            print(f"DEBUG: Saving report mapping to {mapping_path}")
            with open(mapping_path, 'w', encoding='utf-8') as f:
                json.dump(mapping, f, indent=4, ensure_ascii=False)
            return True
        except Exception as e:
            print(f"DEBUG: Save failed: {e}")
            messagebox.showerror("저장 오류", f"매핑 설정을 저장하지 못했습니다: {e}")
            return False

    def open_report_mapping_dialog(self):
        """Open a dialog to configure Excel mapping for Daily Work Report"""
        dialog = tk.Toplevel(self.root)
        dialog.title("작업일보 엑셀 매핑 설정")
        dialog.geometry("600x800")
        dialog.transient(self.root)
        dialog.grab_set()

        mapping = self.load_report_mapping()
        
        main_frame = ttk.Frame(dialog, padding=10)
        main_frame.pack(fill='both', expand=True)

        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill='both', expand=True)

        # Helper to create grid of entries
        def create_entry_grid(parent, fields, current_data):
            entries = {}
            for i, (label_text, key) in enumerate(fields):
                ttk.Label(parent, text=label_text).grid(row=i, column=0, padx=5, pady=5, sticky='w')
                ent = ttk.Entry(parent, width=15)
                # Ensure we handle missing keys or nested dicts in current_data
                val = current_data.get(key, '')
                if isinstance(val, dict): val = val.get('row', '')
                ent.insert(0, str(val))
                ent.grid(row=i, column=1, padx=5, pady=5, sticky='w')
                entries[key] = ent
            return entries

        # Tab 1: General & Header
        tab1 = ttk.Frame(notebook, padding=10)
        notebook.add(tab1, text="일반/상단")
        gen_fields = [
            ("날짜 (Date)", "date_header"),
            ("업체명 (Company)", "company"), ("공사명 (Project)", "project_name"),
            ("적용규격 (Standard)", "standard"), ("장비명 (Equipment)", "equipment"),
            ("성적서번호 (Report No)", "report_no"), ("검사품명 (Item)", "inspection_item"),
            ("검사자 (Inspector)", "inspector"), ("차량번호 (Car No)", "car_no")
        ]
        tab1_data = mapping.get('general', {}).copy()
        tab1_data['date_header'] = mapping.get('header', {}).get('date', 'F2')
        tab1_entries = create_entry_grid(tab1, gen_fields, tab1_data)

        # Tab 2: Methods (Row Numbers)
        tab2 = ttk.Frame(notebook, padding=10)
        notebook.add(tab2, text="검사공법(행)")
        method_keys = ["RT", "UT", "MT", "PT", "HT", "VT", "LT", "ET", "PAUT"]
        method_fields = [(f"{m} (시작 행 번호)", m) for m in method_keys]
        tab2_entries = create_entry_grid(tab2, method_fields, mapping.get('methods', {}))

        # Tab 3: RTK Results
        tab3 = ttk.Frame(notebook, padding=10)
        notebook.add(tab3, text="불량율(RTK)")
        rtk_keys = ['center_miss', 'density', 'marking_miss', 'film_mark', 'handling', 'customer_complaint', 'etc', 'total']
        rtk_fields = [
            ("센터미스", "center_miss"), ("농도", "density"), ("마킹미스", "marking_miss"),
            ("필름마크", "film_mark"), ("취급부주의", "handling"), ("고객불만", "customer_complaint"),
            ("기타", "etc"), ("총계", "total")
        ]
        tab3_entries = create_entry_grid(tab3, rtk_fields, mapping.get('rtk', {}))

        # Tab 4: Materials (Row Numbers)
        tab4 = ttk.Frame(notebook, padding=10)
        notebook.add(tab4, text="자재(행)")
        mat_keys = list(mapping.get('materials', {}).keys())
        if not mat_keys: mat_keys = ['RT T200', 'RT AA400', 'MT WHITE', 'MT 7C-BLACK', 'PT Penetrant', 'PT Cleaner', 'PT Developer']
        mat_fields = [(f"{m} (행 번호)", m) for m in mat_keys]
        tab4_entries = create_entry_grid(tab4, mat_fields, mapping.get('materials', {}))

        def save_and_close():
            try:
                new_mapping = {
                    'header': {'date': tab1_entries['date_header'].get().strip()},
                    'general': {k: tab1_entries[k].get().strip() for k in ['company', 'project_name', 'standard', 'equipment', 'report_no', 'inspection_item', 'inspector', 'car_no']},
                    'methods': {m: {'row': tab2_entries[m].get().strip()} for m in method_keys},
                    'rtk': {k: tab3_entries[k].get().strip() for k in rtk_keys},
                    'ot': mapping.get('ot', {
                        'row1_name': 'B38', 'row1_hours': 'K38',
                        'row2_name': 'B39', 'row2_hours': 'K39'
                    }),
                    'materials': {k: tab4_entries[k].get().strip() for k in mat_keys}
                }
                
                if self.save_report_mapping(new_mapping):
                    messagebox.showinfo("저장 완료", "매핑 설정이 저장되었습니다.")
                    dialog.destroy()
            except Exception as e:
                messagebox.showerror("오류", f"설정 구성 중 오류가 발생했습니다: {e}")

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=10)
        ttk.Button(btn_frame, text="✅ 설정 저장", command=save_and_close).pack(side='right', padx=5)
        ttk.Button(btn_frame, text="❌ 취소", command=dialog.destroy).pack(side='right', padx=5)

    def export_daily_work_report(self, *args, **kwargs):
        if "export" in "export_daily_work_report" or "excel" in "export_daily_work_report":
            from site_apps.central.src.services.excel_exporter import export_daily_work_report_impl
        else:
            from site_apps.central.src.services.data_loader import export_daily_work_report_impl
        return export_daily_work_report_impl(self, *args, **kwargs)

    def export_central_daily_work_report(self, *args, **kwargs):
        from site_apps.central.src.services.excel_exporter import export_central_daily_work_report_impl
        return export_central_daily_work_report_impl(self, *args, **kwargs)

    def _get_usage_session_data(self):
        """Helper to collect Site tab entry data"""
        data = {
            'date': str(self.ent_daily_date.get_date()),
            'site': self.cb_daily_site.get().strip(),
            'company': self.cb_daily_company.get().strip(),
            'applied_code': self.ent_daily_applied_code.get().strip(),
            'report_no': self.ent_daily_report_no.get().strip(),
            'equip': self.cb_daily_equip.get().strip(),
            'material': self.cb_daily_material.get().strip(),
            'inspection_item': self.ent_daily_inspection_item.get().strip(),
            'method': self.cb_daily_test_method.get().strip(),
            'unit': self.cb_daily_unit.get().strip(),
            'amount': self.ent_daily_test_amount.get().strip(),
            'unit_price': self.ent_daily_unit_price.get().strip(),
            'travel_cost': self.ent_daily_travel_cost.get().strip(),
            'meal_cost': self.ent_daily_meal_cost.get().strip(),
            'test_fee': self.ent_daily_test_fee.get().strip(),
            'note': self.ent_daily_note.get().strip(),
            'main_memo': self.main_memo_text.get('1.0', 'end-1c').strip() if hasattr(self, 'main_memo_text') and self.main_memo_text else "",
            'workers': []
        }
        for i in range(1, 11):
            group = getattr(self, f'worker_group{i}', None)
            if group:
                data['workers'].append({
                    'name': group.cb_name.get().strip(),
                    'shift': group.cb_shift.get().strip(),
                    'time': group.ent_worktime.get().strip(),
                    'ot': group.ent_ot.get().strip()
                })
        data['ndt_companies'] = []
        if hasattr(self, 'ndt_company_entries'):
            for entries in self.ndt_company_entries:
                co_data = {'_company': entries['_company'].get().strip()}
                for mat in ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]:
                    co_data[mat] = entries[mat].get().strip()
                data['ndt_companies'].append(co_data)
        data['rtk'] = {cat: ent.get().strip() for cat, ent in self.rtk_entries.items() if cat != '총계'}
        
        # [NEW] Include vehicle inspection data in session snapshot
        data['vehicles'] = []
        if hasattr(self, 'vehicle_boxes'):
            for box in self.vehicle_boxes:
                data['vehicles'].append(box.get_data())
                
        return data

    def _set_usage_session_data(self, data):
        """Helper to restore Site tab entry data"""
        if not data: return
        self.ent_daily_date.set_date(data.get('date', datetime.datetime.now().strftime('%Y-%m-%d')))
        # [FIX] Use the dedicated Company field
        if 'applied_code' in data:
            self.ent_daily_applied_code.delete(0, tk.END)
            self.ent_daily_applied_code.insert(0, data.get('applied_code', ''))
        
        self.cb_daily_company.set(data.get('company', ''))
        self.cb_daily_site.set(data.get('site', ''))
        self.ent_daily_report_no.delete(0, tk.END); self.ent_daily_report_no.insert(0, data.get('report_no', ''))
        if isinstance(self.cb_daily_material, ttk.Combobox):
            self.cb_daily_material.set(data.get('material', ''))
        else:
            self.cb_daily_material.delete(0, tk.END)
            self.cb_daily_material.insert(0, data.get('material', ''))
        if 'inspection_item' in data:
            self.ent_daily_inspection_item.delete(0, tk.END)
            self.ent_daily_inspection_item.insert(0, data.get('inspection_item', ''))
        self.cb_daily_test_method.set(data.get('method', ''))
        self.cb_daily_unit.set(data.get('unit', '매'))
        self.ent_daily_test_amount.delete(0, tk.END); self.ent_daily_test_amount.insert(0, data.get('amount', ''))
        self.ent_daily_unit_price.delete(0, tk.END); self.ent_daily_unit_price.insert(0, data.get('unit_price', ''))
        self.ent_daily_travel_cost.delete(0, tk.END); self.ent_daily_travel_cost.insert(0, data.get('travel_cost', ''))
        self.ent_daily_meal_cost.delete(0, tk.END); self.ent_daily_meal_cost.insert(0, data.get('meal_cost', ''))
        self.ent_daily_test_fee.delete(0, tk.END); self.ent_daily_test_fee.insert(0, data.get('test_fee', ''))
        self.ent_daily_note.delete(0, tk.END); self.ent_daily_note.insert(0, data.get('note', ''))
        if hasattr(self, 'main_memo_text') and self.main_memo_text:
            try:
                self.main_memo_text.delete('1.0', tk.END)
                if 'main_memo' in data and data['main_memo']:
                    self.main_memo_text.insert(tk.END, data['main_memo'])
            except: pass
        workers = data.get('workers', [])
        for i, w_data in enumerate(workers, 1):
            group = getattr(self, f'worker_group{i}', None)
            if group:
                group.cb_name.set(w_data.get('name', ''))
                group.cb_shift.set(w_data.get('shift', ''))
                group.ent_worktime.set(w_data.get('time', ''))
                group.ent_ot.delete(0, tk.END); group.ent_ot.insert(0, w_data.get('ot', ''))
        ndt_data = data.get('ndt_companies', [])
        while len(self.ndt_company_entries) > 1: self.remove_last_ndt_company()
        for i, co_data in enumerate(ndt_data):
            if i >= len(self.ndt_company_entries): self.add_ndt_company_section()
            entries = self.ndt_company_entries[i]
            entries['_company'].set(co_data.get('_company', ''))
            for mat in ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]:
                entries[mat].delete(0, tk.END); entries[mat].insert(0, co_data.get(mat, ''))
        rtk = data.get('rtk', {})
        for cat, val in rtk.items():
            if cat in self.rtk_entries:
                ent = self.rtk_entries[cat]; ent.delete(0, tk.END); ent.insert(0, val)
        self.calculate_rtk_total()
        
        # [NEW] Restore vehicle inspection data
        vehicle_data = data.get('vehicles', [])
        if vehicle_data:
            # Match current boxes or add new ones if needed
            for i, v_state in enumerate(vehicle_data):
                if i < len(self.vehicle_boxes):
                    self.vehicle_boxes[i].set_data(v_state)
                else:
                    # If more boxes are needed than current, we could add them,
                    # but usually there's 1 or the count is already correct.
                    pass

    def _get_inout_session_data(self):
        """Helper to collect In/Out tab entry data"""
        data = {
            'reg': {
                'co_code': self.cb_co_code.get().strip(),
                'eq_code': self.cb_eq_code.get().strip(),
                'item_name': self.cb_item_name.get().strip(),
                'model': self.cb_model.get().strip(),
                'sn': self.ent_sn.get().strip(),
                'class': self.cb_class.get().strip(),
                'spec': self.cb_spec.get().strip(),
                'unit': self.cb_unit.get().strip(),
                'supplier': self.cb_supplier.get().strip(),
                'mfr': self.cb_mfr.get().strip(),
                'origin': self.cb_origin.get().strip(),
                'reorder': self.ent_reorder.get().strip(),
                'init': self.ent_init.get().strip(),
                'price': self.ent_price.get().strip(),
                'cost': self.ent_cost.get().strip()
            },
            'trans': {
                'material': self.cb_material.get().strip(),
                'type': self.cb_type.get().strip(),
                'qty': self.ent_qty.get().strip(),
                'site': self.cb_trans_site.get().strip(),
                'warehouse': self.cb_warehouse.get().strip(),
                'user': self.ent_user.get().strip(),
                'note': self.ent_note.get().strip()
            }
        }
        return data

    def _set_inout_session_data(self, data):
        """Helper to restore In/Out tab entry data"""
        if not data: return
        reg = data.get('reg', {})
        self.cb_co_code.set(reg.get('co_code', ''))
        self.cb_eq_code.set(reg.get('eq_code', ''))
        self.cb_item_name.set(reg.get('item_name', ''))
        self.cb_model.set(reg.get('model', ''))
        self.ent_sn.delete(0, tk.END); self.ent_sn.insert(0, reg.get('sn', ''))
        self.cb_class.set(reg.get('class', ''))
        self.cb_spec.set(reg.get('spec', ''))
        self.cb_unit.set(reg.get('unit', ''))
        self.cb_supplier.set(reg.get('supplier', ''))
        self.cb_mfr.set(reg.get('mfr', ''))
        self.cb_origin.set(reg.get('origin', ''))
        self.ent_reorder.delete(0, tk.END); self.ent_reorder.insert(0, reg.get('reorder', '0'))
        self.ent_init.delete(0, tk.END); self.ent_init.insert(0, reg.get('init', '0'))
        self.ent_price.delete(0, tk.END); self.ent_price.insert(0, reg.get('price', '0'))
        self.ent_cost.delete(0, tk.END); self.ent_cost.insert(0, reg.get('cost', '0'))
        
        trans = data.get('trans', {})
        self.cb_material.set(trans.get('material', ''))
        self.cb_type.set(trans.get('type', 'OUT'))
        self.ent_qty.delete(0, tk.END); self.ent_qty.insert(0, trans.get('qty', ''))
        self.cb_trans_site.set(trans.get('site', ''))
        self.cb_warehouse.set(trans.get('warehouse', ''))
        self.ent_user.set(trans.get('user', ''))
        self.ent_note.delete(0, tk.END); self.ent_note.insert(0, trans.get('note', ''))

    def _get_budget_session_data(self):
        """Helper to collect Budget tab entry data"""
        data = {
            'main': {k: w.get().strip() for k, w in self.budget_widgets.items()},
            'labor_detail': self.labor_detail_widget.get_data(),
            'material_detail': self.material_detail_widget.get_data(),
            'expense_detail': self.expense_detail_widget.get_data()
        }
        return data

    def _set_budget_session_data(self, data):
        """Helper to restore Budget tab entry data"""
        if not data: return
        main = data.get('main', {})
        for k, val in main.items():
            if k in self.budget_widgets:
                w = self.budget_widgets[k]
                if hasattr(w, 'set'): w.set(val)
                else: w.delete(0, tk.END); w.insert(0, val)
        self.labor_detail_widget.set_data(data.get('labor_detail', {}))
        self.material_detail_widget.set_data(data.get('material_detail', {}))
        self.expense_detail_widget.set_data(data.get('expense_detail', {}))
        self._update_budget_kpis()

    def save_form_session(self):
        """Orchestrate saving session data from all entry tabs into one JSON"""
        try:
            global_data = {
                'usage': self._get_usage_session_data(),
                'inout': self._get_inout_session_data(),
                'budget': self._get_budget_session_data(),
                'saved_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            file_path = filedialog.asksaveasfilename(
                title="전체 세션 저장 (일괄)",
                defaultextension=".json",
                filetypes=[("JSON files", "*.json")],
                initialfile=f"PMI_Global_Session_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            )
            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(global_data, f, ensure_ascii=False, indent=2)
                messagebox.showinfo("저장 완료", f"현장/입출고/실행예산 탭의 입력 내용이 일괄 저장되었습니다.\n{os.path.basename(file_path)}")
        except Exception as e:
            messagebox.showerror("오류", f"전체 세션 저장 중 오류 발생: {e}")

    def load_previous_day_data(self):
        """선택된 현장의 가장 최근 기록(전일 데이터)을 불러와 폼을 채웁니다."""
        site = self.cb_daily_site.get().strip()
        if not site:
            messagebox.showwarning("입력 오류", "먼저 현장명을 선택하거나 입력해주세요.")
            return
            
        if self.daily_usage_df.empty:
            messagebox.showinfo("알림", "저장된 기록이 없습니다.")
            return
            
        df_site = self.daily_usage_df[self.daily_usage_df['Site'] == site].copy()
        if df_site.empty:
            messagebox.showinfo("알림", f"'{site}' 현장의 이전 기록이 없습니다.")
            return
            
        # 가장 최근 날짜 찾기
        df_site['Date'] = pd.to_datetime(df_site['Date'])
        recent_date = df_site['Date'].max()
        df_recent = df_site[df_site['Date'] == recent_date]
        
        date_str = recent_date.strftime('%Y-%m-%d')
        if not messagebox.askyesno("전일 데이터 불러오기", f"'{site}' 현장의 가장 최근 기록({date_str})을 불러오시겠습니까?\n(기존 입력된 작업자/차량 정보는 덮어쓰기 됩니다.)"):
            return
            
        try:
            # 1. 작업자 정보 복원 (최초 레코드 기준)
            first_record = df_recent.iloc[0]
            # 0. 일반 정보 복원 (업체명, 장비명 등)
            def _set_val(widget, val):
                if hasattr(widget, 'set'): widget.set(val)
                else:
                    widget.delete(0, 'end')
                    widget.insert(0, val)

            _set_val(self.cb_daily_company, self.clean_nan(first_record.get('업체명', '')))
            _set_val(self.cb_daily_equip, self.clean_nan(first_record.get('장비명', '')))
            
            self.ent_daily_applied_code.delete(0, 'end')
            self.ent_daily_applied_code.insert(0, self.clean_nan(first_record.get('적용코드', '')))
            
            self.ent_daily_inspection_item.delete(0, 'end')
            self.ent_daily_inspection_item.insert(0, self.clean_nan(first_record.get('검사품명', '')))
            
            _set_val(self.cb_daily_test_method, self.clean_nan(first_record.get('검사방법', '')))
            _set_val(self.cb_daily_unit, self.clean_nan(first_record.get('Unit', first_record.get('단위', ''))))
            
            self.ent_daily_unit_price.delete(0, 'end')
            unit_price = self.clean_nan(first_record.get('단가', ''))
            if str(unit_price) in ["0", "0.0"]: unit_price = ""
            self.ent_daily_unit_price.insert(0, unit_price)

            self.ent_daily_report_no.delete(0, 'end')
            self.ent_daily_report_no.insert(0, self.clean_nan(first_record.get('성적서번호', '')))
            
            self.ent_daily_note.delete(0, 'end')
            self.ent_daily_note.insert(0, self.clean_nan(first_record.get('Note', first_record.get('비고', ''))))

            for i in range(1, 11):
                group = getattr(self, f'worker_group{i}', None)
                if group:
                    u_col = 'User' if i == 1 else f'User{i}'
                    t_col = 'WorkTime' if i == 1 else f'WorkTime{i}'
                    o_col = 'OT' if i == 1 else f'OT{i}'
                    m_col = 'Meal' if i == 1 else f'Meal{i}'
                    
                    worker = first_record.get(u_col, first_record.get('작업자' if i == 1 else f'작업자{i}', ''))
                    worktime = first_record.get(t_col, first_record.get('작업시간' if i == 1 else f'작업시간{i}', ''))
                    ot = first_record.get(o_col, first_record.get('OT시간' if i == 1 else f'OT시간{i}', ''))
                    meal = first_record.get(m_col, first_record.get('일비' if i == 1 else f'일비{i}', ''))
                    
                    group.set_worker(self.clean_nan(worker))
                    group.set_time(self.clean_nan(worktime))
                    group.set_ot(self.clean_nan(ot))
                    if hasattr(group, 'set_meal'):
                        group.set_meal(self.clean_nan(meal))
                
            # 2. 차량 정보 복원
            if hasattr(self, 'fixed_vehicle_widget'):
                v_no_str = self.clean_nan(first_record.get('VehicleNo', first_record.get('차량번호', '')))
                v_nos = [x.strip() for x in v_no_str.split("||")] if "||" in v_no_str else [v_no_str]
                
                if v_nos and v_nos[0]:
                    self.fixed_vehicle_widget.cb_vehicle_info.set(v_nos[0])
                    
                if len(v_nos) > 1:
                    existing_boxes = getattr(self, 'vehicle_boxes', [])
                    for idx, extra_v in enumerate(v_nos[1:]):
                        if not extra_v: continue
                        if idx < len(existing_boxes):
                            existing_boxes[idx].cb_vehicle_info.set(extra_v)
                        elif hasattr(self, 'add_vehicle_inspection_box'):
                            new_box = self.add_vehicle_inspection_box()
                            if new_box: new_box.cb_vehicle_info.set(extra_v)
                
            messagebox.showinfo("완료", f"{date_str}의 작업자 및 차량 정보가 성공적으로 불러와졌습니다.\n(검사 물량 및 자재 소모량은 오늘 기준에 맞게 새로 입력해주세요.)")
        except Exception as e:
            messagebox.showerror("오류", f"데이터를 불러오는 중 오류가 발생했습니다: {e}")

    def load_form_session(self):
        """Orchestrate loading session data back to all entry tabs from a JSON file"""
        try:
            file_path = filedialog.askopenfilename(
                title="전체 세션 불러오기 (일괄)",
                filetypes=[("JSON files", "*.json")]
            )
            if not file_path: return
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            # Restore all tabs
            self._set_usage_session_data(data.get('usage'))
            self._set_inout_session_data(data.get('inout'))
            self._set_budget_session_data(data.get('budget'))
            messagebox.showinfo("불러오기 완료", "현장/입출고/실행예산 탭의 데이터가 일괄 복원되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"전체 세션 불러오기 중 오류 발생: {e}")

    def _auto_reconcile_and_register_ndt(self, date_val, site, ndt_data, workers, company_code=""):
        """Unified helper to auto-register missing consumables and create 'OUT' transactions"""
        ndt_product_map = self._load_ndt_product_map()
        ndt_groups = self.ndt_groups
        
        note_suffix = f" ({company_code})" if company_code else ""
        note = f"{site} 현장 사용 (자동 차감){note_suffix}"
        
        # Ensure date_val is datetime and handle date vs datetime objects
        if isinstance(date_val, str):
            date_obj = pd.to_datetime(date_val)
        else:
            date_obj = date_val
            
        # extract date part safely
        pure_date = date_obj.date() if hasattr(date_obj, 'date') else date_obj
        trans_date = datetime.datetime.combine(pure_date, datetime.datetime.now().time())
        
        for name, qty in ndt_data.items():
            if qty <= 0: continue
            
            # 1. Resolve Item Name (Group) and Model
            db_item_name = ndt_product_map.get(name, '')
            db_model_name = name
            if not db_item_name:
                for grp, members in ndt_groups.items():
                    if name in members:
                        db_item_name = grp
                        db_model_name = name # [FIX] Remove prefix for consistency with display logic
                        break
                if not db_item_name: db_item_name = "기타소모품"
            
            # 2. Find or Register (Highly flexible matching for NDT)
            target_item = db_item_name.replace(' ', '')
            
            # [FIX] Fill NaN with empty string before comparison to ensure reliable matching
            m_df = self.materials_df.copy()
            for col in ['품목명', '모델명']:
                if col in m_df.columns:
                    m_df[col] = m_df[col].fillna('').astype(str).str.replace(' ', '', regex=False)

            # Try to find existing item
            # First, check for exact (item+model) match
            target_model = db_model_name.replace(' ', '')
            exact_match = m_df[(m_df['품목명'] == target_item) & (m_df['모델명'] == target_model)]
            
            if not exact_match.empty:
                mat_id = exact_match.iloc[0]['MaterialID']
            else:
                # If no exact match, try matching by item name ONLY for NDT groups
                item_only_match = m_df[m_df['품목명'] == target_item]
                if not item_only_match.empty:
                    # Use the first available item in this group (e.g., generic "PT약품")
                    mat_id = item_only_match.iloc[0]['MaterialID']
                else:
                    # Truly missing - [REFINED] Do NOT register new one automatically
                    print(f"DEBUG: Skipping auto-registration for missing NDT item '{db_item_name}' (Model: {db_model_name})")
                    mat_id = None
            
            # 3. Create Transaction (Only if mat_id exists)
            if mat_id:
                self._create_manual_stock_transaction(trans_date, mat_id, 'OUT', qty, site, workers, note)

    def _create_manual_stock_transaction(self, date_val, mat_id, trans_type, qty, site, user, note):
        # [V23.1_WATERTIGHT_FILTER] Default to block everything unless it matches the whitelist
        try:
            if self.materials_df.empty:
                return # Can't identify item, don't save
                
            # Robust ID matching
            str_target_id = str(mat_id).strip().replace('.0', '')
            # Vectorized lookup for speed
            mask = self.materials_df['MaterialID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True) == str_target_id
            mat_info = self.materials_df[mask]
            
            if mat_info.empty:
                # If we can't find it in DB, we don't know what it is. To be safe, DON'T save it to history.
                return
                
            row = mat_info.iloc[0]
            item_name = str(row.get('품목명', '')).strip()
            model_name = str(row.get('모델명', '')).strip()
            item_name_up = item_name.upper()
            model_name_up = model_name.upper()
            
            # 1. Check for Carestream (Partial Match, Case Insensitive)
            is_carestream = "CARESTREAM" in item_name_up or "CARESTREAM" in model_name_up
            
            # 2. Check for NDT Chemicals (More inclusive partial match for PT/MT/NDT)
            ndt_keywords = ["PT", "MT", "NDT", "침투제", "세척제", "현상제", "자분", "페인트"]
            is_ndt_chem = any(kw in item_name_up or kw in model_name_up for kw in ndt_keywords)
            
            # [CRITICAL] If not a whitelist item, exit NOW
            if not (is_carestream or is_ndt_chem):
                return
                
        except Exception as e:
            print(f"Filter error in _create_manual_stock_transaction: {e}")
            return # Block on error
            
        # --- IF WE GOT HERE, THE ITEM IS ALLOWED ---
        try:
            # Ensure quantity is negative for OUT
            signed_qty = -abs(float(qty)) if trans_type == 'OUT' else abs(float(qty))
            
            new_trans = {
                'Date': pd.to_datetime(date_val),
                'MaterialID': mat_id,
                'Type': trans_type,
                'Quantity': signed_qty,
                'User': user,
                'Site': site,
                'Note': note
            }
            self.transactions_df = pd.concat([self.transactions_df, pd.DataFrame([new_trans])], ignore_index=True)
        except: pass
    
    def refresh_inquiry_filters(self):
        """Unified method to populate inquiry filter dropdowns and sync with autocomplete lists"""
        try:
            import re

            marker_pattern = MARKER_PATTERN
            
            # 1. Collect and Merge Site Info
            raw_sites = set()
            if not self.daily_usage_df.empty and 'Site' in self.daily_usage_df.columns:
                raw_sites.update(self.daily_usage_df['Site'].dropna().astype(str).apply(self.normalize_site_name).tolist())
            if hasattr(self, 'sites'):
                raw_sites.update(self.sites)
            # [NEW] budget_df에 저장된 현장명도 포함
            if hasattr(self, 'budget_df') and not self.budget_df.empty and 'Site' in self.budget_df.columns:
                raw_sites.update(self.budget_df['Site'].dropna().astype(str).apply(self.normalize_site_name).tolist())
            unique_sites = sorted([s for s in raw_sites 
                                     if s and str(s).lower() != 'nan'
                                    and s not in getattr(self, 'hidden_sites', [])])  # [FIX] 숨긴 현장 제외
            
            # [ROBUST] Do NOT overwrite self.sites Master List with history.
            # self.sites is the "Managed List". unique_sites is for "Filters".
            # self.sites[:] = unique_sites # REMOVED: This was causing deleted sites to reappear
            
            if hasattr(self, 'cb_daily_filter_site'):
                self.cb_daily_filter_site['values'] = ['전체'] + unique_sites
                if not self.cb_daily_filter_site.get(): self.cb_daily_filter_site.set('전체')
            if hasattr(self, 'cb_filter_site_monthly'):
                self.cb_filter_site_monthly['values'] = ['전체'] + unique_sites
            # [FIX] 공사실행예산서 상단 현장 콤보박스 → budget_df 현장만 표시
            # (삭제 후 목록에서 즉시 사라지게 하기 위해 daily_usage_df 현장은 제외)
            if hasattr(self, 'cb_budget_site'):
                budget_sites = []
                if hasattr(self, 'budget_df') and not self.budget_df.empty and 'Site' in self.budget_df.columns:
                    budget_sites = sorted([
                        s for s in self.budget_df['Site'].dropna().astype(str).apply(self.normalize_site_name).unique()
                        if s and str(s).lower() != 'nan'
                    ])
                self.budget_sites[:] = budget_sites  # [FIX] 자동완성 목록 동기화
                current_val = self.cb_budget_site.get()
                self.cb_budget_site['values'] = budget_sites
                if current_val: self.cb_budget_site.set(current_val)
            # 하단 실적 조회 콤보박스 → 모든 현장(daily_usage_df + budget_df) 표시
            if hasattr(self, 'cb_budget_view_site'):
                self.cb_budget_view_site['values'] = unique_sites

            # 2. Collect Material Info (History Only)
            raw_materials = set()
            if not self.daily_usage_df.empty and 'MaterialID' in self.daily_usage_df.columns:
                unique_mat_ids = self.daily_usage_df['MaterialID'].dropna().unique()
                for mat_id in unique_mat_ids:
                    name = self.get_material_display_name(mat_id)
                    if name: raw_materials.add(name)
            
            unique_materials = sorted(list(raw_materials))
            
            if hasattr(self, 'cb_daily_filter_material'):
                # [RESTORED] Populate with history names to avoid 'empty' confusion, while still excluding master list
                self.cb_daily_filter_material['values'] = ['전체'] + unique_materials
                if not self.cb_daily_filter_material.get(): self.cb_daily_filter_material.set('전체')
            
            if hasattr(self, 'cb_daily_material'):
                # [NEW] Only update values if it's still a combobox (Site tab uses Entry now)
                if isinstance(self.cb_daily_material, ttk.Combobox):
                    self.cb_daily_material['values'] = unique_materials
            
            if hasattr(self, 'cb_daily_filter_site'):
                self.cb_daily_filter_site['values'] = ['전체'] + unique_sites
            if hasattr(self, 'cb_filter_material_monthly'):
                self.cb_filter_material_monthly['values'] = ['전체'] + unique_materials

            # 3. Equipment dropdown
            if hasattr(self, 'cb_daily_filter_equipment'):
                raw_equip = []
                if not self.daily_usage_df.empty and '장비명' in self.daily_usage_df.columns:
                    raw_equip = self.daily_usage_df['장비명'].dropna().astype(str).str.strip().unique().tolist()
                
                # Combine with catalog if needed
                unique_equip = sorted(list(set([e for e in raw_equip if e and str(e).lower() != 'nan'])))
                self.cb_daily_filter_equipment['values'] = ['전체'] + unique_equip
                if not self.cb_daily_filter_equipment.get(): self.cb_daily_filter_equipment.set('전체')

            # 4. Worker dropdown
            if hasattr(self, 'cb_daily_filter_worker'):
                worker_cols = ['User', 'User2', 'User3', 'User4', 'User5', 'User6', 'User7', 'User8', 'User9', 'User10']
                all_workers_raw = set()
                if not self.daily_usage_df.empty:
                    for col in worker_cols:
                        if col in self.daily_usage_df.columns:
                            all_workers_raw.update(self.daily_usage_df[col].dropna().astype(str).unique().tolist())
                
                # Include static list
                if hasattr(self, 'users'): all_workers_raw.update(self.users)
                
                clean_workers_set = set()
                for w in all_workers_raw:
                    w_str = str(w).strip()
                    if not w_str or w_str.lower() in ['nan', '0.0', 'none']: continue
                    cleaned = marker_pattern.sub('', w_str).strip()
                    cleaned = " ".join(cleaned.split())
                    if cleaned: clean_workers_set.add(cleaned)

                unique_workers = sorted(list(clean_workers_set))
                # self.users[:] = unique_workers # REMOVED: This was causing deleted workers to reappear
                self.cb_daily_filter_worker['values'] = ['전체'] + unique_workers
                if not self.cb_daily_filter_worker.get(): self.cb_daily_filter_worker.set('전체')

            # 5. Vehicle dropdown
            raw_vehicles = set()
            if not self.daily_usage_df.empty and '차량번호' in self.daily_usage_df.columns:
                v_list = self.daily_usage_df['차량번호'].dropna().astype(str).str.strip().tolist()
                for v in v_list:
                    cleaned = self.clean_nan(v)
                    if cleaned: raw_vehicles.add(cleaned)
            if hasattr(self, 'vehicles'):
                raw_vehicles.update(self.vehicles)
                
            unique_vehicles = sorted(list([v for v in raw_vehicles if v and str(v).lower() != 'nan']))
            if hasattr(self, 'cb_daily_filter_vehicle'):
                self.cb_daily_filter_vehicle['values'] = ['전체'] + unique_vehicles
                if not self.cb_daily_filter_vehicle.get(): self.cb_daily_filter_vehicle.set('전체')
                
            # 6. Company dropdown (for entry form)
            raw_companies = set()
            if not self.daily_usage_df.empty and '업체명' in self.daily_usage_df.columns:
                raw_companies.update(self.daily_usage_df['업체명'].dropna().astype(str).str.strip().tolist())
            if hasattr(self, 'companies'):
                raw_companies.update(self.companies)
                
            unique_companies = sorted(list([c for c in raw_companies if c and str(c).lower() != 'nan']))
            if hasattr(self, 'cb_daily_company'):
                self.cb_daily_company['values'] = unique_companies
            if hasattr(self, 'cb_daily_filter_company'):
                self.cb_daily_filter_company['values'] = ['전체'] + unique_companies
                if not self.cb_daily_filter_company.get(): self.cb_daily_filter_company.set('전체')
            if hasattr(self, 'cb_daily_filter_vehicle'):
                self.cb_daily_filter_vehicle['values'] = ['전체'] + unique_vehicles
                if not self.cb_daily_filter_vehicle.get(): self.cb_daily_filter_vehicle.set('전체')
        except Exception as e:
            print(f"ERROR in refresh_inquiry_filters: {e}")
        if hasattr(self, 'cb_sales_filter_site'):
            self.cb_sales_filter_site['values'] = ['전체'] + self.sites
            if not self.cb_sales_filter_site.get(): self.cb_sales_filter_site.set('전체')
        if hasattr(self, 'cb_trans_filter_vehicle'):
            self.cb_trans_filter_vehicle['values'] = ['전체'] + unique_vehicles

    def update_daily_usage_view(self, *args, **kwargs):
        from site_apps.central.src.views.daily_usage_view import update_daily_usage_view_impl
        return update_daily_usage_view_impl(self, *args, **kwargs)

    def reset_daily_usage_filters(self):
        """Reset all daily usage history filters to default values"""
        try:
            # 1. Reset Dates
            start_date = datetime.datetime(2024, 1, 1)
            self.ent_daily_start_date.set_date(start_date)
            self.ent_daily_end_date.set_date(datetime.datetime.now())
            
            # 2. Reset Comboboxes
            filter_combos = [
                self.cb_daily_filter_company, self.cb_daily_filter_site, 
                self.cb_daily_filter_material, self.cb_daily_filter_equipment,
                self.cb_daily_filter_worker, self.cb_daily_filter_vehicle,
                self.cb_daily_filter_shift
            ]
            for combo in filter_combos:
                if hasattr(self, combo.winfo_name()) or True: # Safeguard
                    combo.set('전체')
            
            # 3. Refresh View
            self.update_daily_usage_view()
            
        except Exception as e:
            print(f"Error resetting daily usage filters: {e}")

    def _auto_adjust_tree_columns(self, tree, expand_only=False):
        """Automatically adjust column widths to fit content"""
        import tkinter.font as tkfont
        
        # Use the actual font used in the Treeview (12pt Malgun Gothic)
        # This ensures measurement matches display perfectly.
        content_font = tkfont.Font(family="Malgun Gothic", size=12)
        heading_font = tkfont.Font(family="Malgun Gothic", size=12, weight="bold")
        
        # We only care about visible columns
        visible_cols = tree['displaycolumns']
        if not visible_cols or visible_cols == ('#all'):
             visible_cols = tree['columns']

        for col in visible_cols:
            # Maximum Overkill: Character Count Heuristic
            # Assume 40px per character + 200px buffer
            # This GUARANTEES visibility in any environment
            
            # Measure heading
            w = len(col) * 40 + 200
            
            # Measure content
            col_index = list(tree['columns']).index(col)
            for item in tree.get_children():
                val = tree.item(item, 'values')
                if val and col_index < len(val):
                    text_len = len(str(val[col_index]))
                    content_w = text_len * 40 + 200
                    if content_w > w:
                        w = content_w
            
            # Apply specific constraints
            # Ultra-Safe: No max limit for potentially long text columns
            if col in ['품목명', '작업자', '비고']:
                w = max(w, 300) # Minimum base
                if col == '품목명': w = max(w, 700)
                elif col == '작업자': w = max(w, 300)
                elif col == '비고': w = max(w, 600)
            else:
                # Other columns get a relaxed global max
                min_w = 200
                if col == '날짜': min_w = 400
                elif col == '현장': min_w = 300
                elif col == '장비명': min_w = 250
                elif col == '검사방법': min_w = 200
                
                w = max(min_w, min(w, 2000)) 
            
            # Smart Auto-Expand: Only grow if content requires it
            if expand_only:
                current_w = int(tree.column(col, 'width'))
                if w <= current_w:
                    continue  # Keep user's manual adjustment
            
            # All columns are now user-resizable
            tree.column(col, width=w, minwidth=20, stretch=False, anchor='center')
        
        # Force layout update
        tree.update_idletasks()

    def delete_daily_usage_entry(self):
        """선택된 일일 사용량 기록 삭제 (인덱스 기반으로 정확하게 삭제 및 재고 환원)"""
        selected_items = self.daily_usage_tree.selection()
        
        if not selected_items:
            messagebox.showwarning("선택 오류", "삭제할 항목을 선택해주세요.")
            return
        
        # 사용자 확인
        result = messagebox.askyesno("삭제 확인", f"{len(selected_items)}개의 기록을 삭제하시겠습니까?\n(삭제 시 차감되었던 재고도 자동으로 환원됩니다.)")
        
        if not result:
            return
        
        indices_to_delete = []
        for item in selected_items:
            tags = self.daily_usage_tree.item(item, 'tags')
            if tags:
                try:
                    df_idx = int(tags[0])
                    indices_to_delete.append(df_idx)
                except ValueError:
                    continue
        
        if not indices_to_delete:
            messagebox.showwarning("삭제 실패", "삭제할 항목의 데이터 정보를 찾을 수 없습니다.")
            return

        deleted_count = 0
        try:
            valid_indices_to_delete = []
            for idx in indices_to_delete:
                if idx not in self.daily_usage_df.index:
                    continue
                
                valid_indices_to_delete.append(idx)
                entry = self.daily_usage_df.loc[idx]
                
                # --- [FIX] 재고 환원 로직 (자동 차감된 트랜잭션만 선택적으로 환원) ---
                site = entry.get('Site', '')
                usage_date = pd.to_datetime(entry.get('Date'))
                mat_id_to_delete = entry.get('MaterialID', '')
                
                if not self.transactions_df.empty:
                    # 해당 현장/날짜의 "(자동 차감)" 트랜잭션을 찾아 삭제
                    trans_mask = (
                        (pd.to_datetime(self.transactions_df['Date'], errors='coerce').dt.normalize() == pd.to_datetime(usage_date).normalize()) &
                        (self.transactions_df['Site'].astype(str) == str(site)) &
                        (self.transactions_df['MaterialID'].astype(str) == str(mat_id_to_delete)) &
                        (self.transactions_df['Type'] == 'OUT') &
                        (self.transactions_df['Note'].str.contains(f"{site} 현장 사용", na=False, regex=False))
                    )
                    self.transactions_df = self.transactions_df[~trans_mask]
                    deleted_count += 1

            if valid_indices_to_delete:
                # 기록 삭제
                self.daily_usage_df = self.daily_usage_df.drop(valid_indices_to_delete)
            
            # 인덱스 초기화는 하지 않음 (태그 매칭을 위해 원본 인덱스 유지 권장하나, 
            # drop 후에는 뷰 업데이트 시 어차피 다시 생성되므로 안전하게 reset_index 가능)
            self.daily_usage_df = self.daily_usage_df.reset_index(drop=True)
            
            if self.save_data():
                messagebox.showinfo("삭제 완료", f"{len(indices_to_delete)}개의 기록이 삭제되고 재고가 환원되었습니다.")
                self.update_daily_usage_view()
                self.update_recent_entries_view() # Ensure the mini table is updated!
                self.update_stock_view()
                self.update_transaction_view()
                self.refresh_inquiry_filters()
                
                # [FIX] Automatically refresh the Site tab and Query tab
                if hasattr(self, 'budget_view_tree'):
                    self.update_budget_site_view()
                if hasattr(self, 'query_tree') and hasattr(self, 'cb_filter_year'):
                    try:
                        y = int(self.cb_filter_year.get().replace('년', ''))
                        m = int(self.cb_filter_month.get().replace('월', ''))
                        self.update_monthly_usage_view(y, m)
                    except:
                        pass
        except Exception as e:
            messagebox.showerror("삭제 오류", f"기록 삭제 중 오류가 발생했습니다: {e}")

    def open_edit_daily_usage_dialog(self):
        """Open a dialog to edit the selected daily usage record"""
        selection = self.daily_usage_tree.selection()
        if not selection:
            messagebox.showwarning("선택 오류", "수정할 항목을 선택해주세요.")
            return
            
        item = self.daily_usage_tree.item(selection[0])
        tags = item.get('tags', [])
        if not tags:
            messagebox.showwarning("데이터 오류", "선택한 항목의 데이터 정보를 찾을 수 없습니다.")
            return
            
        try:
            df_idx = int(tags[0])
            entry_data = self.daily_usage_df.loc[df_idx]
        except (ValueError, KeyError, IndexError):
            messagebox.showwarning("데이터 오류", "데이터를 불러오는 중 오류가 발생했습니다.")
            return

        # Create Edit Dialog
        edit_win = tk.Toplevel(self.root)
        edit_win.title("일일 사용 기록 수정")
        edit_win.geometry("800x900")
        edit_win.transient(self.root)
        edit_win.grab_set()

        # Main scrollable frame
        container = ttk.Frame(edit_win)
        container.pack(expand=True, fill='both')
        
        canvas = tk.Canvas(container)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas, padding=20)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Scroll binding is handled globally in MaterialManager.__init__

        fields = {}
        
        # 1. Basic Info
        basic_frame = ttk.LabelFrame(scrollable_frame, text="기본 정보", padding=10)
        basic_frame.pack(fill='x', pady=5)
        
        # Date
        ttk.Label(basic_frame, text="날짜:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        ent_date = DateEntry(basic_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        curr_date = entry_data.get('Date', datetime.datetime.now().strftime('%Y-%m-%d'))
        try: ent_date.set_date(pd.to_datetime(curr_date))
        except: pass
        ent_date.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        fields['Date'] = ent_date

        # Site
        ttk.Label(basic_frame, text="현장:").grid(row=0, column=2, padx=5, pady=5, sticky='w')
        cb_site = ttk.Combobox(basic_frame, width=20, values=self.sites)
        cb_site.set(self.clean_nan(entry_data.get('Site', '')))
        cb_site.grid(row=0, column=3, padx=5, pady=5, sticky='w')
        fields['Site'] = cb_site
        
        # Company & Applied Code
        ttk.Label(basic_frame, text="업체명:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
        cb_company = ttk.Combobox(basic_frame, width=20, values=self.companies)
        cb_company.set(self.clean_nan(entry_data.get('업체명', '')))
        cb_company.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        fields['업체명'] = cb_company
        
        ttk.Label(basic_frame, text="적용코드:").grid(row=1, column=2, padx=5, pady=5, sticky='w')
        ent_code = ttk.Entry(basic_frame, width=12)
        ent_code.insert(0, self.clean_nan(entry_data.get('적용코드', '')))
        ent_code.grid(row=1, column=3, padx=5, pady=5, sticky='w')
        fields['적용코드'] = ent_code

        # Material
        ttk.Label(basic_frame, text="품목명:").grid(row=2, column=0, padx=5, pady=5, sticky='w')
        cb_mat = ttk.Combobox(basic_frame, width=40)
        mat_options = []
        for _, m_row in self.materials_df.iterrows():
            if m_row.get('Active', 1) == 0: continue
            mat_options.append(self.get_material_display_name(m_row['MaterialID']))
        
        # Deduplicate and sort for a clean list
        mat_options = sorted(list(set([m for m in mat_options if m])))
        cb_mat['values'] = mat_options
        
        # Determine current display name
        curr_mat_id = entry_data.get('MaterialID')
        curr_mat_display = self.get_material_display_name(curr_mat_id)
        
        # [FIX] Use raw MaterialID string if not found in master data (e.g. JIREH Scanner)
        if not curr_mat_display and curr_mat_id and not pd.isna(curr_mat_id):
            if str(curr_mat_id).strip().lower() not in ('nan', ''):
                curr_mat_display = str(curr_mat_id).strip()
        
        cb_mat.set(curr_mat_display)
        cb_mat.grid(row=2, column=1, columnspan=3, padx=5, pady=5, sticky='w')
        fields['Material'] = cb_mat

        # Equipment & Method
        ttk.Label(basic_frame, text="장비명:").grid(row=3, column=0, padx=5, pady=5, sticky='w')
        cb_equip = ttk.Combobox(basic_frame, width=20, values=self.equipments)
        cb_equip.set(self.clean_nan(entry_data.get('장비명', '')))
        cb_equip.grid(row=3, column=1, padx=5, pady=5, sticky='w')
        fields['장비명'] = cb_equip

        ttk.Label(basic_frame, text="검사방법:").grid(row=3, column=2, padx=5, pady=5, sticky='w')
        cb_method = ttk.Combobox(basic_frame, width=10, values=['RT', 'PAUT', 'UT', 'MT', 'PT', 'ETC'])
        cb_method.set(self.clean_nan(entry_data.get('검사방법', '')))
        cb_method.grid(row=3, column=3, padx=5, pady=5, sticky='w')
        fields['검사방법'] = cb_method
        
        # [V13_FIX] Add Unit to basic info - Now using dynamic list from unit management
        ttk.Label(basic_frame, text="단위:").grid(row=3, column=4, padx=5, pady=5, sticky='w')
        cb_unit = ttk.Combobox(basic_frame, width=10, values=self.daily_units)
        cb_unit.set(self.clean_nan(entry_data.get('Unit', entry_data.get('단위', '매'))))
        cb_unit.grid(row=3, column=5, padx=5, pady=5, sticky='w')
        fields['Unit'] = cb_unit
        
        # Inspection Item
        ttk.Label(basic_frame, text="검사품명:").grid(row=4, column=0, padx=5, pady=5, sticky='w')
        ent_item = ttk.Entry(basic_frame, width=40)
        ent_item.insert(0, self.clean_nan(entry_data.get('검사품명', '')))
        ent_item.grid(row=4, column=1, columnspan=3, padx=5, pady=5, sticky='w')
        fields['검사품명'] = ent_item

        # 2. Quantities & Costs
        cost_frame = ttk.LabelFrame(scrollable_frame, text="수량 및 비용", padding=10)
        cost_frame.pack(fill='x', pady=5)
        
        qty_configs = [
            ('검사량', '검사량'), ('단가', '단가'), ('출장비', '출장비'),
            ('일식', '일식'), ('검사비', '검사비'), ('필름매수', 'FilmCount')
        ]
        for i, (lbl, key) in enumerate(qty_configs):
            row = i // 3
            col = (i % 3) * 2
            ttk.Label(cost_frame, text=f"{lbl}:").grid(row=row, column=col, padx=5, pady=5, sticky='w')
            ent = ttk.Entry(cost_frame, width=12)
            val = entry_data.get(key, 0)
            if pd.isna(val): val = 0
            ent.insert(0, str(val))
            ent.grid(row=row, column=col+1, padx=5, pady=5, sticky='w')
            fields[key] = ent

        # 3. RTK & NDT
        material_frame = ttk.Frame(scrollable_frame)
        material_frame.pack(fill='x', pady=5)
        
        rtk_frame = ttk.LabelFrame(material_frame, text="RTK 사용량 (재촬영)", padding=10)
        rtk_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))
        
        rtk_cats = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타"]
        for i, cat in enumerate(rtk_cats):
            ttk.Label(rtk_frame, text=f"{cat}:").grid(row=i, column=0, padx=5, pady=2, sticky='w')
            ent = ttk.Entry(rtk_frame, width=10)
            val = entry_data.get(f'RTK_{cat}', 0)
            if pd.isna(val): val = 0
            ent.insert(0, str(val))
            ent.grid(row=i, column=1, padx=5, pady=2, sticky='w')
            fields[f'RTK_{cat}'] = ent

        ndt_frame = ttk.LabelFrame(material_frame, text="NDT 약품 사용량", padding=10)
        ndt_frame.pack(side='left', fill='both', expand=True, padx=(5, 0))
        
        ndt_cats = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
        for i, cat in enumerate(ndt_cats):
            ttk.Label(ndt_frame, text=f"{cat}:").grid(row=i, column=0, padx=5, pady=2, sticky='w')
            ent = ttk.Entry(ndt_frame, width=10)
            m_key = "".join(cat.split())
            val = entry_data.get(f'NDT_{m_key}', 0)
            if pd.isna(val): val = 0
            ent.insert(0, str(val))
            ent.grid(row=i, column=1, padx=5, pady=2, sticky='w')
            fields[f'NDT_{m_key}'] = ent

            fields[f'NDT_{m_key}'] = ent
            
        # 4. NDT 상세 조건
        ndt_cond_frame = ttk.LabelFrame(scrollable_frame, text="NDT 상세 조건", padding=10)
        ndt_cond_frame.pack(fill='x', pady=5)
        
        cond_configs = [
            ('작업형태', '작업형태'), ('조건1', '조건1'), ('조건2', '조건2'),
            ('제경비율(%)', '제경비율'), ('기술료율(%)', '기술료율'), ('보정계수', '보정계수'),
            ('환산물량', '환산물량'), ('제경비', '제경비'), ('기술료', '기술료')
        ]
        for c_i, (c_lbl, c_key) in enumerate(cond_configs):
            c_row = c_i // 3
            c_col = (c_i % 3) * 2
            ttk.Label(ndt_cond_frame, text=f"{c_lbl}:").grid(row=c_row, column=c_col, padx=5, pady=2, sticky='w')
            c_ent = ttk.Entry(ndt_cond_frame, width=15)
            c_val = entry_data.get(c_key, '')
            if pd.isna(c_val): c_val = ''
            c_ent.insert(0, self.clean_nan(c_val))
            c_ent.grid(row=c_row, column=c_col+1, padx=5, pady=2, sticky='w')
            fields[c_key] = c_ent

        # 5. Vehicle Info
        vehicle_frame = ttk.LabelFrame(scrollable_frame, text="차량 점검 정보", padding=10)
        vehicle_frame.pack(fill='x', pady=5)
        
        v_configs = [
            ('차량번호', '차량번호'), ('주행거리', '주행거리'),
            ('점검내용', '차량점검'), ('차량비고', '차량비고')
        ]
        for v_i, (v_lbl, v_key) in enumerate(v_configs):
            v_row = v_i // 2
            v_col = (v_i % 2) * 2
            ttk.Label(vehicle_frame, text=f"{v_lbl}:").grid(row=v_row, column=v_col, padx=5, pady=5, sticky='w')
            v_ent = ttk.Entry(vehicle_frame, width=25)
            v_val = entry_data.get(v_key, '')
            v_ent.insert(0, self.clean_nan(v_val))
            v_ent.grid(row=v_row, column=v_col+1, padx=5, pady=5, sticky='w')
            fields[v_key] = v_ent

        # [NEW] 5.5 Remark (비고/메모)
        remark_frame = ttk.Frame(scrollable_frame, padding=10)
        remark_frame.pack(fill='x', pady=5)
        ttk.Label(remark_frame, text="비고(메모):").pack(side='left', padx=5)
        ent_remark = ttk.Entry(remark_frame, width=80)
        curr_rem = entry_data.get('Note', entry_data.get('비고', ''))
        if pd.isna(curr_rem): curr_rem = ''
        ent_remark.insert(0, self.clean_nan(curr_rem))
        ent_remark.pack(side='left', fill='x', expand=True, padx=5)
        fields['Note'] = ent_remark

        # 6. Workers
        worker_main_frame = ttk.LabelFrame(scrollable_frame, text="작업자 정보 (1~10)", padding=10)
        worker_main_frame.pack(fill='x', pady=5)
        


        # [NEW] Inner-dialog helper for Fee calculation
        def update_edit_fee_calc(event=None):
            try:
                def get_f(val):
                    try:
                        v = str(val).strip().replace(',', '')
                        return float(v) if v else 0.0
                    except: return 0.0

                qty = get_f(fields['검사량'].get())
                price = get_f(fields['단가'].get())
                travel = get_f(fields['출장비'].get())
                meal = get_f(fields['일식'].get())
                
                calc_fee = (qty * price) + travel + meal
                fields['검사비'].delete(0, tk.END)
                fields['검사비'].insert(0, f"{calc_fee:,.0f}")
            except: pass

        # Bindings for auto-calc
        fields['검사량'].bind('<KeyRelease>', lambda e: update_edit_fee_calc())
        
        # [NEW] Add comma auto-formatting for edit fields
        for k in ['단가', '출장비', '일식', '검사비']:
            if k in fields:
                fields[k].bind('<KeyRelease>', update_edit_fee_calc, add='+')
                fields[k].bind('<FocusOut>', lambda e, widget=fields[k]: self.format_entry_with_commas(e, widget), add='+')
                fields[k].bind('<Return>', lambda e, widget=fields[k]: self.format_entry_with_commas(e, widget), add='+')
        
        # Ensure FilmCount field also formats nicely (integer)
        if 'FilmCount' in fields:
            fields['FilmCount'].bind('<FocusOut>', lambda e, widget=fields['FilmCount']: self.format_entry_with_commas(e, widget), add='+')

        # Grid for workers: 5 rows x 2 cols
        worker_fields = {}
        for i in range(1, 11):
            w_row = (i-1) // 2
            w_col = (i-1) % 2
            
            w_frame = ttk.Frame(worker_main_frame, padding=2)
            w_frame.grid(row=w_row, column=w_col, padx=10, pady=5, sticky='nw')
            
            ttk.Label(w_frame, text=f"{i}:", width=2).pack(side='left')
            
            # Name
            cb_name = ttk.Combobox(w_frame, width=8, values=[''] + self.users)
            name_key = 'User' if i == 1 else f'User{i}'
            cb_name.set(self.clean_nan(entry_data.get(name_key, '')))
            cb_name.pack(side='left', padx=2)
            worker_fields[f'name{i}'] = cb_name
            
            # Time
            cb_time = ttk.Combobox(w_frame, width=16, values=[''] + self.worktimes)
            time_key = 'WorkTime' if i == 1 else f'WorkTime{i}'
            cb_time.set(self.clean_nan(entry_data.get(time_key, '')))
            cb_time.pack(side='left', padx=2)
            worker_fields[f'time{i}'] = cb_time
            
            # OT
            ent_ot = ttk.Entry(w_frame, width=18)
            ot_key = 'OT' if i == 1 else f'OT{i}'
            ent_ot.insert(0, self.clean_nan(entry_data.get(ot_key, '')))
            ent_ot.pack(side='left', padx=2)
            worker_fields[f'ot{i}'] = ent_ot
            
            # Auto-calculate OT for each worker when time changes
            def on_time_changed(event, idx=i, w_ent=cb_time, ot_ent=ent_ot, d_ent=ent_date):
                self.calculate_and_update_ot_manual(w_ent.get(), ot_ent, d_ent.get_date())
            
            cb_time.bind('<Return>', on_time_changed)
        
        # [NEW] Apply autocomplete to main edit fields
        self.enable_autocomplete(cb_site, values_list_attr='sites')
        self.enable_autocomplete(cb_mat, values_list=mat_options)
        self.enable_autocomplete(cb_equip, values_list_attr='equipment_suggestions')

        # 5. Buttons
        btn_frame = ttk.Frame(scrollable_frame, padding=20)
        btn_frame.pack(fill='x')
        
        def save_edits():
            # Collect data
            new_data = {}
            new_data['Date'] = ent_date.get_date().strftime('%Y-%m-%d')
            new_data['Site'] = cb_site.get().strip()
            new_data['업체명'] = cb_company.get().strip()
            new_data['적용코드'] = ent_code.get().strip()
            new_data['검사품명'] = ent_item.get().strip()
            
            full_mat_name = cb_mat.get().strip()
            # [FIX] Robust MaterialID mapping: Try to find matching ID using standard display name generator
            found_id = None
            if full_mat_name:
                for _, m_row in self.materials_df.iterrows():
                    m_id_check = m_row['MaterialID']
                    if self.get_material_display_name(m_id_check) == full_mat_name:
                        found_id = m_id_check
                        break
                
                # [FIX] Only auto-register if it's a CONSUMABLE. Equipment should NOT be registered.
                if not found_id:
                    if self._is_consumable_material(full_mat_name, cb_method.get().strip()):
                        found_id = self.register_new_material(full_mat_name, warehouse='현장', 규격='수정등록')
                    else:
                        found_id = full_mat_name
            
            new_data['Material'] = full_mat_name
            new_data['MaterialID'] = found_id if found_id is not None else entry_data.get('MaterialID')
            
            new_data['장비명'] = cb_equip.get().strip()
            new_data['검사방법'] = cb_method.get().strip()
            new_data['Unit'] = cb_unit.get().strip() # [V13_FIX] Save Unit field
            
            # Numeric fields
            for key in ['검사량', '단가', '출장비', '일식', '검사비', 'FilmCount']:
                try: 
                    v_str = fields[key].get().strip().replace(',', '')
                    new_data[key] = float(v_str) if v_str else 0.0
                except: new_data[key] = 0.0
                
            for cat in rtk_cats:
                try: new_data[f'RTK_{cat}'] = float(fields[f'RTK_{cat}'].get())
                except: new_data[f'RTK_{cat}'] = 0.0
                
            for cat in ndt_cats:
                m_key = "".join(cat.split())
                try: new_data[f'NDT_{m_key}'] = float(fields[f'NDT_{m_key}'].get())
                except: new_data[f'NDT_{m_key}'] = 0.0
                
            # String fields
            for key in ['차량번호', '주행거리', '차량점검', '차량비고', '작업형태', '조건1', '조건2', 'Note']:
                if key in fields:
                    new_data[key] = fields[key].get().strip()
            
            # Additional numeric fields
            for key in ['제경비율', '기술료율', '보정계수', '환산물량', '제경비', '기술료']:
                if key in fields:
                    try:
                        v_str = fields[key].get().strip().replace(',', '')
                        new_data[key] = float(v_str) if v_str else 0.0
                    except: new_data[key] = 0.0
                
            for i in range(1, 11):
                name_key = 'User' if i == 1 else f'User{i}'
                time_key = 'WorkTime' if i == 1 else f'WorkTime{i}'
                ot_key = 'OT' if i == 1 else f'OT{i}'
                
                new_data[name_key] = worker_fields[f'name{i}'].get().strip()
                new_data[time_key] = worker_fields[f'time{i}'].get().strip()
                new_data[ot_key] = worker_fields[f'ot{i}'].get().strip()

            # Provide immediate feedback
            btn_save_edit.config(state='disabled', text="저장 중...")
            edit_win.update_idletasks()
            
            try:
                if self.save_daily_usage_edits(df_idx, new_data):
                    edit_win.destroy()
                    messagebox.showinfo("수정 완료", "기록이 성공적으로 수정되었으며 재고 내역이 업데이트되었습니다.")
                else:
                    btn_save_edit.config(state='normal', text="수정사항 저장")
            except Exception as e:
                btn_save_edit.config(state='normal', text="수정사항 저장")
                messagebox.showerror("오류", f"수정 중 오류가 발생했습니다: {e}")

        btn_save_edit = ttk.Button(btn_frame, text="수정사항 저장", style='Big.TButton', command=save_edits)
        btn_save_edit.pack(side='right', padx=10)
        ttk.Button(btn_frame, text="취소", command=edit_win.destroy).pack(side='right', padx=10)

    def _calculate_ot_from_worktime(self, *args, **kwargs):
        from site_apps.central.src.models.worker_model import _calculate_ot_from_worktime_impl
        return _calculate_ot_from_worktime_impl(self, *args, **kwargs)

    def calculate_and_update_ot_manual(self, worktime_value, ot_entry, calculation_date=None):
        """Helper for personal calculation during editing without complex widget lookups"""
        try:
            current_date = calculation_date if calculation_date else self.ent_daily_date.get_date()
            ot_hours, amount = self._calculate_ot_from_worktime(worktime_value, current_date)
            
            if amount > 0:
                ot_entry.delete(0, tk.END)
                ot_entry.insert(0, f"{amount:,}")
            elif ot_hours == 0:
                # If no OT, clear it or set to 0
                ot_entry.delete(0, tk.END)
                ot_entry.insert(0, "0")
        except: pass

    def _safe_set_daily_df(self, idx, col, val, is_numeric=False):
        """Defensively set a value in daily_usage_df handling all possible dtype conflicts"""
        if col not in self.daily_usage_df.columns:
            self.daily_usage_df.at[idx, col] = val
            return

        col_dtype = self.daily_usage_df[col].dtype
        is_stringy = pd.api.types.is_string_dtype(col_dtype) or pd.api.types.is_object_dtype(col_dtype)
        
        v_to_set = val
        if is_stringy and not is_numeric:
            # Force to cleaned string for string columns
            v_to_set = str(val).strip() if pd.notna(val) else ""
            if v_to_set in ('0.0', '0', 'nan', 'None', 'NaT'): v_to_set = ""
        elif is_numeric:
            # Force to float for numeric columns
            try:
                if isinstance(val, str):
                    v_to_set = float(val.replace(',', '').strip()) if val.strip() else 0.0
                else:
                    v_to_set = float(val) if pd.notna(val) else 0.0
            except:
                v_to_set = 0.0

        try:
            self.daily_usage_df.at[idx, col] = v_to_set
        except:
            try:
                # Type conflict (e.g. float into strict string column or vice versa)
                # Cast the column to object to allow the mixed/new type
                self.daily_usage_df[col] = self.daily_usage_df[col].astype(object)
                self.daily_usage_df.at[idx, col] = v_to_set
            except:
                # Final fallback
                self.daily_usage_df.at[idx, col] = str(v_to_set)

    def _is_consumable_material(self, *args, **kwargs):
        from site_apps.central.src.models.material_model import _is_consumable_material_impl
        return _is_consumable_material_impl(self, *args, **kwargs)

    def save_daily_usage_edits(self, df_idx, new_data):
        """Save edited daily usage and reconcile stock"""
        try:
            old_entry = self.daily_usage_df.loc[df_idx]
            
            # 1. Revert Old Stock (Delete existing auto-deductions before applying new ones)
            old_site = old_entry.get('Site', '')
            old_date = pd.to_datetime(old_entry.get('Date'))
            
            if not self.transactions_df.empty:
                trans_mask = (
                    (pd.to_datetime(self.transactions_df['Date'], errors='coerce').dt.normalize() == pd.to_datetime(old_date).normalize()) &
                    (self.transactions_df['Site'] == old_site) &
                    (self.transactions_df['Type'] == 'OUT') &
                    (self.transactions_df['Note'].str.contains(f"{old_site} 현장 사용", na=False))
                )
                self.transactions_df = self.transactions_df[~trans_mask]

            # 2. Update entry in DF
            # [FIX] Handle column name aliases (Date/날짜, Site/현장) to ensure existing columns are updated
            site_col = 'Site' if 'Site' in self.daily_usage_df.columns else '현장'
            date_col = 'Date' if 'Date' in self.daily_usage_df.columns else '날짜'
            
            # [NEW] Pre-collect existing column names to avoid duplicate-like column creation
            existing_cols = self.daily_usage_df.columns
            
            for k, v in new_data.items():
                target_key = k
                if k == 'Site': target_key = site_col
                elif k == 'Date': target_key = date_col
                
                # [FIX] For NDT and RTK fields, ensure they match normalized column names in the DF
                if k.startswith('NDT_') or k.startswith('RTK_'):
                    # Search for case-insensitive and space-insensitive match in existing columns
                    k_norm = k.replace(' ', '').upper()
                    for ex_col in existing_cols:
                        if str(ex_col).replace(' ', '').upper() == k_norm:
                            target_key = ex_col
                            break
                
                # Comprehensive list of columns that MUST be numeric
                numeric_cols = ['Usage', '수량', '검사량', '단가', '출장비', '일식', '검사비', 'FilmCount', 'OT', 'OT금액']
                is_numeric = k.startswith('NDT_') or k.startswith('RTK_') or k in numeric_cols or any(f"OT{i}" == k or f"OT금액{i}" == k for i in range(1, 11))
                
                # Use safe setter to avoid all dtype crashes
                self._safe_set_daily_df(df_idx, target_key, v, is_numeric=(is_numeric or k == 'MaterialID'))

            # Ensure Usage is consistently updated from Inspection Amount (검사량)
            usage_col = 'Usage' if 'Usage' in self.daily_usage_df.columns else '수량'
            self._safe_set_daily_df(df_idx, usage_col, new_data.get('검사량', 0.0), is_numeric=True)
            if '검사량' in self.daily_usage_df.columns:
                self._safe_set_daily_df(df_idx, '검사량', new_data.get('검사량', 0.0), is_numeric=True)
            
            # [FIX] Initialize new_mat_display for stock reconciliation logic
            new_mat_display = new_data.get('Material', '')
            if not new_mat_display:
                new_mat_display = self.get_material_display_name(new_data.get('MaterialID', ''))
            
            # 3. Apply New Deduction (Selective Auto-deduction)
            new_date = pd.to_datetime(new_data['Date'])
            new_site = new_data['Site']
            # [NEW] PAUT 및 장비류는 재고 차감에서 제외
            method_raw = new_data.get('검사방법', '')
            method = str(method_raw).strip().upper()
            is_excluded = (method == 'PAUT')
            
            if not is_excluded:
                # Resolve or Register MaterialID
                new_mat_id = ""
                if new_mat_display:
                    for _, row in self.materials_df.iterrows():
                        if self.get_material_display_name(row['MaterialID']) == new_mat_display:
                            new_mat_id = row['MaterialID']
                            if row.get('Active', 1) == 0:
                                self.materials_df.loc[self.materials_df['MaterialID'] == new_mat_id, 'Active'] = 1
                            break
                    
                    if not new_mat_id and new_mat_display:
                        if self._is_consumable_material(new_mat_display, method):
                            # [UX IMPROVEMENT] 등록되지 않은 소모품 감지 시 승인 팝업 추가 (수정/저장 동일)
                            if messagebox.askyesno("신규 자재 등록", f"등록되지 않은 소모품 자재명('{new_mat_display}')입니다.\n재고 마스터에 신규 자재로 등록하시겠습니까?\n\n(아니오를 누르면 기록에는 변경되지만 재고 연동(차감) 대상에서 제외됩니다.)"):
                                new_mat_id = self.register_new_material(new_mat_display, warehouse='현장', 규격='자동등록')
                            else:
                                new_mat_id = new_mat_display
                        else:
                            new_mat_id = new_mat_display
                
                new_qty = float(new_data.get('검사량', 0))
                new_note_pattern = f'{new_site} 현장 사용 (자동 차감)'
                
                workers_names = [new_data.get(f'User{i}' if i > 1 else 'User', '').strip() for i in range(1, 11)]
                all_workers = ", ".join([n for n in workers_names if n])
                
                if new_mat_id and new_qty > 0:
                    if self._is_consumable_material(new_mat_display, method):
                        mat_info = self.get_material_info(new_mat_id)
                        full_item_name = str(mat_info.get('품목명', '')).replace(' ', '').upper()
                        # PT/MT 약품 부모 항목 중복 차감 방지
                        if full_item_name not in ["PT약품", "MT약품", "NDT약품"]:
                            self._create_manual_stock_transaction(new_date, new_mat_id, 'OUT', new_qty, new_site, all_workers, new_note_pattern)
                
                # NDT 약품(소모품) 일괄 정산
                ndt_data = {
                    name: float(new_data.get(f'NDT_{"".join(name.split())}', 0))
                    for name in ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
                }
                self._auto_reconcile_and_register_ndt(new_date, new_site, ndt_data, all_workers, new_data.get('회사코드', ''))

            # 4. Save & Refresh
            if self.save_data():
                self.update_daily_usage_view()
                self.refresh_inquiry_filters()
                self.update_stock_view()
                self.update_transaction_view()
                self.update_material_combo() # Refresh history suggestions
                return True
            return False
        except Exception as e:
            messagebox.showerror("수정 오류", f"기록 수정 중 오류가 발생했습니다: {e}")
            return False
                
    def export_daily_usage_history(self):
        """Export current filtered daily usage history to Excel"""
        try:
            # Get selected items only, or all if none selected
            selected_items = self.daily_usage_tree.selection()
            if not selected_items:
                # If no items selected, export all visible items
                items_to_export = self.daily_usage_tree.get_children()
            else:
                # Export only selected items
                items_to_export = selected_items
            
            data = []
            for item in items_to_export:
                data.append(self.daily_usage_tree.item(item, 'values'))
            
            if not data:
                messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
                return
            
            # Export columns currently visible in the Treeview
            display_cols = self.daily_usage_tree['displaycolumns']
            if not display_cols or display_cols == ('#all'):
                selected_cols = list(self.daily_usage_tree['columns'])
            else:
                selected_cols = list(display_cols)
            
            # Map selected column names to their indices in the Treeview values
            all_cols = list(self.daily_usage_tree['columns'])
            col_indices = [all_cols.index(col) for col in selected_cols]
            
            # Reconstruct data
            filtered_data = []
            for item_values in data:
                row = [item_values[i] for i in col_indices if i < len(item_values)]
                filtered_data.append(row)
            
            filename = f"일일사용내역_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=filename, filetypes=[("Excel files", "*.xlsx")])
            
            if save_path:
                cols = list(self.daily_usage_tree['columns'])
                col_to_idx = {col: i for i, col in enumerate(cols)}
                
                # Rebuild data from 'data' (which has ALL columns including hidden ones)
                # and swap summarized '작업자' with full list from '(Full작업자)'
                final_filtered_data = []
                for item_values in data:
                    row_list = list(item_values)
                    
                    # Perform the swap if both columns exist
                    if '작업자' in cols and '(Full작업자)' in col_to_idx:
                        full_idx = col_to_idx['(Full작업자)']
                        if full_idx < len(row_list):
                            full_val = row_list[full_idx]
                            if full_val:
                                # Find index of '작업자' in ALL columns to swap in row_list
                                idx_worker = cols.index('작업자')
                                row_list[idx_worker] = full_val
                    
                    # Extract only the selected display columns for final export
                    final_row = [row_list[i] for i in col_indices if i < len(row_list)]
                    final_filtered_data.append(final_row)
                
                df = pd.DataFrame(final_filtered_data, columns=selected_cols)
                df = self.clean_df_export(df)
                self.save_df_to_excel_autofit(df, save_path, "일일사용내역")
                messagebox.showinfo("완료", f"데이터가 엑셀로 저장되었습니다.\n{save_path}")
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 내보내기 중 오류가 발생했습니다: {e}")

    def show_column_visibility_dialog(self, tree=None):
        """Open dialog to manually show/hide columns in the specified tree view"""
        if tree is None:
            tree = self.daily_usage_tree
            
        all_cols = list(tree['columns'])
        
        # Mandatory columns identification
        # NDT/RTK columns are at the end, workers in middle, basic info at start
        selectable_cols = [c for c in all_cols if c != '(Full작업자)']
        
        # [FIX] Use stored manual preferences instead of current filtered displaycolumns
        # so that hidden-by-data columns still appear checked if user wants them.
        if tree == self.daily_usage_tree:
            active_cols = getattr(self, 'manual_visible_cols', [])
            if not active_cols:
                # Default to all non-data columns if no preference saved
                active_cols = [c for c in all_cols if c != '(Full작업자)']
        else:
            active_cols = tree['displaycolumns']
            if not active_cols or active_cols == ('#all'):
                 active_cols = all_cols
        
        dialog = ColumnSelectionDialog(self.root, selectable_cols, title="표시 컬럼 관리")
        # Overwrite vars with current visibility state
        for col, var in dialog.vars.items():
            var.set(col in active_cols)
            
        dialog.wait_window()
        
        if dialog.result is not None:
            # Result set from dialog
            final_selection = list(dialog.result)
            
            # Reconstruct final display based on user result + mandatory ones
            mandatory_lookup = {
                self.daily_usage_tree: ['날짜', '업체명', '현장', '작업자', '검사품명', '수량', '단위', '단가', '검사비', 'OT시간', 'OT금액', '비고'],
                self.inout_tree: ['날짜', '구분', '품목명']
            }
            
            mandatory_cols = mandatory_lookup.get(tree, ['날짜'])
            for mc in mandatory_cols:
                if mc in all_cols and mc not in final_selection:
                    final_selection.append(mc)
            
            # Re-order to match original tree columns
            sorted_selection = [c for c in all_cols if c in final_selection]
            
            tree['displaycolumns'] = sorted_selection
            
            # Save configuration to the appropriate key
            if tree == self.daily_usage_tree:
                self.manual_visible_cols = sorted_selection
                if not hasattr(self, 'tab_config'): self.tab_config = {}
                self.tab_config['daily_usage_visible_cols'] = sorted_selection
            elif tree == self.inout_tree:
                if not hasattr(self, 'tab_config'): self.tab_config = {}
                self.tab_config['inout_visible_cols'] = sorted_selection
                
            self.save_tab_config()
            
            # Immediate refresh to apply selection
            if tree == self.daily_usage_tree:
                self.update_daily_usage_view()
            elif tree == self.inout_tree:
                self.update_transaction_view()

    def export_weekly_report(self):
        """주간 업무보고 엑셀(알아서 생성) 출력"""
        import pandas as pd
        import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import os
        
        top = tk.Toplevel(self.root)
        top.title("주간 업무보고서 출력")
        top.geometry("400x350")
        top.transient(self.root)
        top.grab_set()
        
        ttk.Label(top, text="🗓️ 주간 업무보고서 자동 생성", font=('Arial', 12, 'bold')).pack(pady=10)
        
        now = datetime.datetime.now()
        monday = now - datetime.timedelta(days=now.weekday())
        sunday = monday + datetime.timedelta(days=6)
        
        frame1 = ttk.Frame(top)
        frame1.pack(pady=10)
        
        from tkcalendar import DateEntry

        ttk.Label(frame1, text="시작일 (월):").grid(row=0, column=0, padx=5, pady=5)
        start_cal = DateEntry(frame1, width=12, background='darkblue',
                              foreground='white', borderwidth=2,
                              date_pattern='yyyy-mm-dd', year=monday.year, month=monday.month, day=monday.day)
        start_cal.grid(row=0, column=1, padx=5)
        
        ttk.Label(frame1, text="종료일 (일):").grid(row=1, column=0, padx=5, pady=5)
        end_cal = DateEntry(frame1, width=12, background='darkblue',
                            foreground='white', borderwidth=2,
                            date_pattern='yyyy-mm-dd', year=sunday.year, month=sunday.month, day=sunday.day)
        end_cal.grid(row=1, column=1, padx=5)
        
        frame2 = ttk.Frame(top)
        frame2.pack(pady=5, fill='both', expand=True, padx=10)
        ttk.Label(frame2, text="다음 주 작업 예정 및 의견:").pack(anchor='w')
        next_week_txt = tk.Text(frame2, height=4, width=40)
        next_week_txt.pack(fill='both', expand=True, pady=5)
        
        def do_export():
            s_date = start_cal.get().strip()
            e_date = end_cal.get().strip()
            next_plan = next_week_txt.get("1.0", "end-1c").strip()
            try:
                s_dt = pd.to_datetime(s_date).date()
                e_dt = pd.to_datetime(e_date).date()
            except Exception:
                messagebox.showerror("오류", "날짜 형식이 올바르지 않습니다 (YYYY-MM-DD)")
                return
                
            history_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'daily_work_history.json')
            if not os.path.exists(history_path):
                messagebox.showinfo("알림", "저장된 작업일보 데이터가 없습니다.")
                return
                
            import json
            with open(history_path, 'r', encoding='utf-8') as f:
                history = json.load(f)
                
            week_data = {}
            cumulative_data = {}
            for d_str, data in history.items():
                try:
                    d_obj = pd.to_datetime(d_str).date()
                    if s_dt <= d_obj <= e_dt:
                        week_data[d_str] = data
                    if d_obj <= e_dt:
                        cumulative_data[d_str] = data
                except: pass

            if not cumulative_data:
                messagebox.showinfo("알림", f"{e_date}까지 저장된 작업일보 데이터가 없습니다.")
                return

            def get_personnel_count(data, use_report_default=False):
                """작업일보의 검사원 인원을 안전하게 숫자로 변환한다."""
                p_data = data.get('personnel_data', {})
                count = 0
                try:
                    count_str = str(p_data.get('검사원_누계', '0')).strip()
                    if count_str and count_str.isdigit() and count_str != '0':
                        count = int(count_str)
                    if count == 0:
                        inspector = str(p_data.get('검사원_인원', '0')).strip()
                        manager = str(p_data.get('검사원_현장대리인', '0')).strip()
                        count = (int(inspector) if inspector.isdigit() else 0) + (int(manager) if manager.isdigit() else 0)
                except Exception:
                    count = 0
                return count if count > 0 else (1 if use_report_default else 0)

            cumulative_time = 0
            cumulative_ndt_methods = {}
            for data in cumulative_data.values():
                cumulative_time += get_personnel_count(data) * 8
                for result in data.get('ndt_results', []):
                    method = str(result.get("검사방법", "")).strip().upper()
                    if method:
                        cumulative_ndt_methods[method] = cumulative_ndt_methods.get(method, 0) + 1
                
            wb = Workbook()
            ws = wb.active
            ws.title = "주간업무보고"
            
            title_font = Font(name='맑은 고딕', size=16, bold=True)
            head_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
            norm_font = Font(name='맑은 고딕', size=10)
            head_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            center_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            ws.merge_cells('A1:G1')
            ws['A1'] = f"주간 업무보고서 ({s_date} ~ {e_date})"
            ws['A1'].font = title_font
            ws['A1'].alignment = center_align
            
            headers = ['일자', '현장명', '작업내용', '투입인원', '작업시간', '검사실적', '비고']
            for col_num, head in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = head
                cell.font = head_font
                cell.fill = head_fill
                cell.alignment = center_align
                cell.border = thin_border
                
            row_idx = 4
            total_time = 0
            total_ndt_methods = {}
            center_align_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Sort dates
            sorted_dates = sorted(week_data.keys())

            if not sorted_dates:
                ws.cell(row=row_idx, column=1, value=f"{s_date} ~ {e_date}").alignment = center_align
                ws.cell(row=row_idx, column=2, value="중앙지사 관내").alignment = center_align
                ws.cell(row=row_idx, column=3, value="금주 신규 작업 없음").alignment = center_align
                ws.cell(row=row_idx, column=4, value="-").alignment = center_align
                ws.cell(row=row_idx, column=5, value="0 시간").alignment = center_align
                ws.cell(row=row_idx, column=6, value="0 POINT").alignment = center_align
                ws.cell(row=row_idx, column=7, value="누계 및 차주 계획 참조").alignment = center_align_wrap
                for col_num in range(1, 8):
                    ws.cell(row=row_idx, column=col_num).border = thin_border
                    ws.cell(row=row_idx, column=col_num).font = norm_font
                row_idx += 1
            
            for date_val in sorted_dates:
                data = week_data[date_val]
                
                # NDT results
                ndt_results = data.get('ndt_results', [])
                daily_ndt_methods = {}
                for r in ndt_results:
                    method = str(r.get("검사방법", "")).strip().upper()
                    if method:
                        daily_ndt_methods[method] = daily_ndt_methods.get(method, 0) + 1
                        total_ndt_methods[method] = total_ndt_methods.get(method, 0) + 1
                
                daily_ndt_texts = [f"{m} {c} POINT" for m, c in daily_ndt_methods.items()]
                ndt_text = "\n".join(daily_ndt_texts) if daily_ndt_texts else "0 POINT"
                
                # Personnel data
                w_count = get_personnel_count(data, use_report_default=True)
                    
                total_time += w_count * 8
                
                ws.cell(row=row_idx, column=1, value=str(date_val)).alignment = center_align
                ws.cell(row=row_idx, column=2, value="중앙지사 관내").alignment = center_align
                ws.cell(row=row_idx, column=3, value=f"{w_count}명 작업 진행").alignment = center_align
                ws.cell(row=row_idx, column=4, value=f"{w_count} 명").alignment = center_align
                ws.cell(row=row_idx, column=5, value=f"{w_count * 8} 시간").alignment = center_align
                ws.cell(row=row_idx, column=6, value=ndt_text).alignment = center_align_wrap
                ws.cell(row=row_idx, column=7, value="").alignment = center_align
                
                for col_num in range(1, 8):
                    ws.cell(row=row_idx, column=col_num).border = thin_border
                    ws.cell(row=row_idx, column=col_num).font = norm_font
                    
                row_idx += 1
                
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 22
            ws.column_dimensions['G'].width = 20
            
            ws.merge_cells(f'A{row_idx}:C{row_idx}')
            ws.cell(row=row_idx, column=1, value="주간 합계").alignment = center_align
            ws.cell(row=row_idx, column=1).font = Font(name='맑은 고딕', size=11, bold=True)
            for col_num in range(1, 4):
                ws.cell(row=row_idx, column=col_num).border = thin_border
                
            ws.cell(row=row_idx, column=4, value=f"-").alignment = center_align
            ws.cell(row=row_idx, column=5, value=f"{total_time} 시간").alignment = center_align
            
            total_ndt_texts = [f"{m} {c} POINT" for m, c in total_ndt_methods.items()]
            total_ndt_text = "\n".join(total_ndt_texts) if total_ndt_texts else "0 POINT"
            ws.cell(row=row_idx, column=6, value=total_ndt_text).alignment = center_align_wrap
            ws.cell(row=row_idx, column=7, value="").alignment = center_align
            
            for col_num in range(4, 8):
                ws.cell(row=row_idx, column=col_num).border = thin_border
                ws.cell(row=row_idx, column=col_num).font = Font(name='맑은 고딕', size=11, bold=True)

            if not week_data:
                row_idx += 2
                cumulative_ndt_texts = [f"{m} {c} POINT" for m, c in sorted(cumulative_ndt_methods.items())]
                cumulative_ndt_text = ", ".join(cumulative_ndt_texts) if cumulative_ndt_texts else "검사실적 없음"
                ws.merge_cells(f'A{row_idx}:G{row_idx+2}')
                cell = ws.cell(row=row_idx, column=1)
                cell.value = (f"■ 작업 누계 (최초 작업일 ~ {e_date})\n"
                              f"작업일보 {len(cumulative_data)}일 / 투입 {cumulative_time} 시간 / {cumulative_ndt_text}")
                cell.font = Font(name='맑은 고딕', size=11)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                for r in range(row_idx, row_idx+3):
                    for c in range(1, 8):
                        ws.cell(row=r, column=c).border = thin_border

            row_idx += 4 if not week_data else 2
            ws.merge_cells(f'A{row_idx}:G{row_idx+2}')
            cell = ws.cell(row=row_idx, column=1)
            cell.value = f"■ 다음 주 작업 예정 및 의견\n{next_plan or '미입력'}"
            cell.font = Font(name='맑은 고딕', size=11)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Draw borders for the merged cell block
            for r in range(row_idx, row_idx+3):
                for c in range(1, 8):
                    ws.cell(row=r, column=c).border = thin_border
                        
            out_name = f"주간업무보고_{s_date.replace('-','')}_{e_date.replace('-','')}.xlsx"
            out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), out_name)
            
            try:
                wb.save(out_path)
                
                msg_extra = ""
                try:
                    import win32com.client as win32
                    import time
                    excel = win32.gencache.EnsureDispatch('Excel.Application')
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    
                    wb_com = excel.Workbooks.Open(out_path)
                    ws_com = wb_com.Sheets(1)
                    ws_com.UsedRange.CopyPicture(Format=2)
                    time.sleep(0.5)
                    wb_com.Close(False)
                    excel.Quit()
                    
                    msg_extra = "\n\n💡 꿀팁: 카카오톡 전송을 위해 표가 '이미지'로 복사되었습니다!\n(카톡 창에서 Ctrl+V 를 누르시면 표가 그대로 붙여넣기 됩니다)"
                except Exception as ex:
                    msg_extra = f"\n(클립보드 이미지 복사 실패: {ex})"
                    
                messagebox.showinfo("완료", f"주간보고서가 생성되었습니다.\n{out_path}{msg_extra}")
                os.startfile(out_path)
                top.destroy()
            except Exception as e:
                messagebox.showerror("오류", f"파일 저장 실패: {e}")
                
        ttk.Button(top, text="보고서 생성 및 열기", command=do_export, style='Accent.TButton' if 'Accent.TButton' in self.style.theme_names() else 'TButton').pack(pady=20)

    def export_monthly_ndt_report(self):
        """일일 사용 데이터를 월용역진도보고서 '3. 비파괴검사 현황' 시트에 자동 채움"""
        import copy
        
        if self.daily_usage_df.empty:
            messagebox.showinfo("알림", "저장된 작업일보 데이터가 없습니다.")
            return

        # --- 0. 설정 다이얼로그 ---
        top = tk.Toplevel(self.root)
        top.title("월간 진도보고서 - 비파괴검사 현황 내보내기")
        top.geometry("550x500")
        top.transient(self.root)
        top.grab_set()
        
        ttk.Label(top, text="📊 월간 진도보고서 비파괴검사 현황 자동 입력", 
                  font=('Arial', 12, 'bold')).pack(pady=10)
        
        # 대상 기간 선택
        period_frame = ttk.LabelFrame(top, text="1. 대상 기간 (당월)")
        period_frame.pack(fill='x', padx=15, pady=5)
        
        now = datetime.datetime.now()
        ttk.Label(period_frame, text="연도:").pack(side='left', padx=5)
        year_var = tk.IntVar(value=now.year)
        ttk.Spinbox(period_frame, from_=2024, to=2030, textvariable=year_var, width=6).pack(side='left')
        ttk.Label(period_frame, text="  월:").pack(side='left', padx=5)
        month_var = tk.IntVar(value=now.month)
        ttk.Spinbox(period_frame, from_=1, to=12, textvariable=month_var, width=4).pack(side='left')
        
        ttk.Label(period_frame, text="  문서번호:").pack(side='left', padx=5)
        doc_num_var = tk.StringVar(value="01")
        ttk.Entry(period_frame, textvariable=doc_num_var, width=5).pack(side='left')
        
        # 현장 → 열배관/관리소 매핑
        map_frame = ttk.LabelFrame(top, text="2. 현장명 → 시트 매핑 (열배관 / 관리소 구분)")
        map_frame.pack(fill='x', padx=15, pady=5)
        
        # 현재 데이터에 존재하는 현장명(Site) 목록 추출
        site_list = sorted(self.daily_usage_df['Site'].dropna().unique().tolist()) if 'Site' in self.daily_usage_df.columns else []
        
        ttk.Label(map_frame, text="열배관 현장명 (쉼표 구분):").pack(anchor='w', padx=5, pady=2)
        main_var = tk.StringVar(value="")
        ttk.Entry(map_frame, textvariable=main_var, width=60).pack(padx=5, pady=2)
        
        ttk.Label(map_frame, text="관리소 현장명 (쉼표 구분):").pack(anchor='w', padx=5, pady=2)
        mgmt_var = tk.StringVar(value="")
        ttk.Entry(map_frame, textvariable=mgmt_var, width=60).pack(padx=5, pady=2)
        
        ttk.Label(map_frame, text=f"💡 현재 등록된 현장: {', '.join(site_list) if site_list else '(없음)'}", 
                  foreground='blue', wraplength=500).pack(padx=5, pady=5)
        
        # 파일 선택
        file_frame = ttk.LabelFrame(top, text="3. 대상 엑셀 파일")
        file_frame.pack(fill='x', padx=15, pady=5)
        
        file_var = tk.StringVar(value=r"C:\Users\jjch2\Desktop\월용역진도보고서.xlsx")
        ttk.Entry(file_frame, textvariable=file_var, width=55).pack(side='left', padx=5, pady=5)
        def browse_file():
            fp = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")], title="월용역진도보고서 선택")
            if fp: file_var.set(fp)
        ttk.Button(file_frame, text="찾기", command=browse_file).pack(side='left', padx=5)
        
        # 진행
        log_text = tk.Text(top, height=8, state='disabled', wrap='word')
        log_text.pack(fill='both', expand=True, padx=15, pady=5)
        
        def log(msg):
            log_text.config(state='normal')
            log_text.insert(tk.END, msg + "\n")
            log_text.see(tk.END)
            log_text.config(state='disabled')
            top.update()
        
        def do_export():
            try:
                year = year_var.get()
                month = month_var.get()
                doc_num = doc_num_var.get().strip() or "01"
                filepath = file_var.get().strip()
                
                main_sites = [s.strip() for s in main_var.get().split(',') if s.strip()]
                mgmt_sites = [s.strip() for s in mgmt_var.get().split(',') if s.strip()]
                
                if not main_sites and not mgmt_sites:
                    messagebox.showwarning("입력 오류", "열배관 또는 관리소에 해당하는 현장명을 최소 1개 입력해주세요.")
                    return
                if not filepath:
                    messagebox.showwarning("입력 오류", "대상 엑셀 파일을 선택해주세요.")
                    return
                
                log(f"▶ 기간: {year}년 {month}월")
                log(f"▶ 열배관 현장: {main_sites}")
                log(f"▶ 관리소 현장: {mgmt_sites}")
                
                # --- 1. 데이터 필터링 ---
                df = self.daily_usage_df.copy()
                df['_date'] = pd.to_datetime(df['Date'], errors='coerce')
                df = df[(df['_date'].dt.year == year) & (df['_date'].dt.month == month)]
                log(f"▶ {year}년 {month}월 데이터: {len(df)}건")
                
                if df.empty:
                    messagebox.showinfo("결과", f"{year}년 {month}월 데이터가 없습니다.")
                    return
                
                # --- 2. 관경 → 필름규격 매핑 ---
                def get_film_type(inch_val):
                    """관경(Inch)에서 RT 필름 규격 결정"""
                    try:
                        v = str(inch_val).replace('"', '').replace("'", '').strip()
                        if '/' in v and len(v) <= 5:  # 3/4 등 분수
                            parts = v.split('/')
                            num = float(parts[0]) / float(parts[1])
                        else:
                            num = float(v)
                        
                        if num >= 20:
                            return 'B'   # B-TYPE (3⅓"×17")
                        elif num >= 6:
                            return 'A'   # A-TYPE (3⅓"×12")
                        else:
                            return 'A/2' # A/2-TYPE (3⅓"×6")
                    except:
                        return 'A'  # 기본값
                
                # --- 3. 집계 함수 ---
                def aggregate_site(site_df):
                    """한 현장(열배관 or 관리소)의 데이터를 집계"""
                    result = {
                        'RT': {'B': {}, 'A': {}, 'A/2': {}},
                        'UT': {'data': {}},
                        'PT': {'data': {}}
                    }
                    # 각 셀의 키: (shift, insp_type)
                    # shift: '주간' or '야간/휴일'
                    # insp_type: 'ORI' or 'REP'
                    
                    for _, row in site_df.iterrows():
                        method = str(row.get('검사방법', '')).strip().upper()
                        if method not in ['RT', 'UT', 'PT', 'PAUT']:
                            continue
                        
                        work_type = str(row.get('작업형태', '주간')).strip()
                        if work_type in ['야간', '휴일', '야간/휴일']:
                            shift = '야간/휴일'
                        else:
                            shift = '주간'
                        
                        insp_type = str(row.get('검사구분', 'ORI')).strip().upper()
                        if insp_type not in ['ORI', 'REP']:
                            insp_type = 'ORI'
                        
                        try:
                            qty = float(str(row.get('검사량', row.get('Usage', 0))).replace(',', '') or 0)
                        except:
                            qty = 0.0
                        
                        try:
                            joints = float(str(row.get('조인트수', 0)).replace(',', '') or 0)
                        except:
                            joints = 0.0
                        
                        try:
                            corr_factor = float(str(row.get('보정계수', 1)).replace(',', '') or 1)
                        except:
                            corr_factor = 1.0
                        
                        try:
                            adj_qty = float(str(row.get('환산물량', 0)).replace(',', '') or 0)
                        except:
                            adj_qty = qty * corr_factor
                        
                        key = (shift, insp_type)
                        
                        if method == 'RT':
                            film_type = get_film_type(row.get('관경(Inch)', ''))
                            bucket = result['RT'][film_type]
                            if key not in bucket:
                                bucket[key] = {'qty': 0, 'joints': 0}
                            bucket[key]['qty'] += qty
                            bucket[key]['joints'] += joints
                        
                        elif method == 'UT':
                            bucket = result['UT']['data']
                            if key not in bucket:
                                bucket[key] = {'qty': 0, 'joints': 0, 'adj_qty': 0}
                            bucket[key]['qty'] += qty
                            bucket[key]['joints'] += joints
                            bucket[key]['adj_qty'] += adj_qty
                        
                        elif method == 'PT':
                            bucket = result['PT']['data']
                            if key not in bucket:
                                bucket[key] = {'qty': 0, 'joints': 0, 'adj_qty': 0}
                            bucket[key]['qty'] += qty
                            bucket[key]['joints'] += joints
                            bucket[key]['adj_qty'] += adj_qty
                    
                    return result
                
                # --- 4. 열배관 / 관리소 각각 집계 ---
                main_df = df[df['Site'].isin(main_sites)] if main_sites else pd.DataFrame()
                mgmt_df = df[df['Site'].isin(mgmt_sites)] if mgmt_sites else pd.DataFrame()
                
                log(f"▶ 열배관 데이터: {len(main_df)}건, 관리소 데이터: {len(mgmt_df)}건")
                
                main_agg = aggregate_site(main_df) if not main_df.empty else None
                mgmt_agg = aggregate_site(mgmt_df) if not mgmt_df.empty else None
                
                                # --- 5. 엑셀 기입 (openpyxl로 변경) ---
                import openpyxl
                from openpyxl.utils import get_column_letter
                import os
                
                save_path = filepath
                wb = openpyxl.load_workbook(save_path)
                
                def write_ndt_sheet_openpyxl(ws, agg):
                    if agg is None: return
                    
                    def get_val(bucket, shift, insp_type, field='qty'):
                        key = (shift, insp_type)
                        return bucket.get(key, {}).get(field, 0)
                    
                    def safe_set(row, col, val):
                        if val and val != 0:
                            ws.cell(row=row, column=col).value = round(val, 2) if isinstance(val, float) and val != int(val) else int(val) if isinstance(val, float) and val == int(val) else val
                            
                    rt_types = {'B': 8, 'A': 9, 'A/2': 10}
                    rt_total = {'주간_joints': 0, '주간_ORI': 0, '주간_REP': 0,
                                '야간_joints': 0, '야간_ORI': 0, '야간_REP': 0}
                    
                    for film_key, row_num in rt_types.items():
                        bucket = agg['RT'][film_key]
                        j_day = get_val(bucket, '주간', 'ORI', 'joints') + get_val(bucket, '주간', 'REP', 'joints')
                        ori_day = get_val(bucket, '주간', 'ORI', 'qty')
                        rep_day = get_val(bucket, '주간', 'REP', 'qty')
                        sum_day = ori_day + rep_day
                        j_night = get_val(bucket, '야간/휴일', 'ORI', 'joints') + get_val(bucket, '야간/휴일', 'REP', 'joints')
                        ori_night = get_val(bucket, '야간/휴일', 'ORI', 'qty')
                        rep_night = get_val(bucket, '야간/휴일', 'REP', 'qty')
                        sum_night = ori_night + rep_night
                        j_total = j_day + j_night
                        ori_total = ori_day + ori_night
                        rep_total = rep_day + rep_night
                        sum_total = ori_total + rep_total
                        
                        safe_set(row_num, 3, j_day)
                        safe_set(row_num, 4, ori_day)
                        safe_set(row_num, 5, rep_day)
                        safe_set(row_num, 6, sum_day)
                        safe_set(row_num, 7, j_night)
                        safe_set(row_num, 8, ori_night)
                        safe_set(row_num, 9, rep_night)
                        safe_set(row_num, 10, sum_night)
                        safe_set(row_num, 11, j_total)
                        safe_set(row_num, 12, ori_total)
                        safe_set(row_num, 13, rep_total)
                        safe_set(row_num, 14, sum_total)
                        
                        rt_total['주간_joints'] += j_day
                        rt_total['주간_ORI'] += ori_day
                        rt_total['주간_REP'] += rep_day
                        rt_total['야간_joints'] += j_night
                        rt_total['야간_ORI'] += ori_night
                        rt_total['야간_REP'] += rep_night
                        
                    safe_set(11, 3, rt_total['주간_joints'])
                    safe_set(11, 4, rt_total['주간_ORI'])
                    safe_set(11, 5, rt_total['주간_REP'])
                    safe_set(11, 6, rt_total['주간_ORI'] + rt_total['주간_REP'])
                    safe_set(11, 7, rt_total['야간_joints'])
                    safe_set(11, 8, rt_total['야간_ORI'])
                    safe_set(11, 9, rt_total['야간_REP'])
                    safe_set(11, 10, rt_total['야간_ORI'] + rt_total['야간_REP'])
                    safe_set(11, 11, rt_total['주간_joints'] + rt_total['야간_joints'])
                    safe_set(11, 12, rt_total['주간_ORI'] + rt_total['야간_ORI'])
                    safe_set(11, 13, rt_total['주간_REP'] + rt_total['야간_REP'])
                    safe_set(11, 14, rt_total['주간_ORI'] + rt_total['주간_REP'] + rt_total['야간_ORI'] + rt_total['야간_REP'])
                    
                    ut = agg['UT']['data']
                    ut_day_j = get_val(ut, '주간', 'ORI', 'joints') + get_val(ut, '주간', 'REP', 'joints')
                    ut_day_ori = get_val(ut, '주간', 'ORI', 'qty')
                    ut_day_rep = get_val(ut, '주간', 'REP', 'qty')
                    ut_night_j = get_val(ut, '야간/휴일', 'ORI', 'joints') + get_val(ut, '야간/휴일', 'REP', 'joints')
                    ut_night_ori = get_val(ut, '야간/휴일', 'ORI', 'qty')
                    ut_night_rep = get_val(ut, '야간/휴일', 'REP', 'qty')
                    
                    safe_set(12, 3, ut_day_j)
                    safe_set(12, 4, ut_day_ori + ut_day_rep)
                    safe_set(12, 6, ut_day_ori + ut_day_rep)
                    safe_set(12, 7, ut_night_j)
                    safe_set(12, 8, ut_night_ori + ut_night_rep)
                    safe_set(12, 10, ut_night_ori + ut_night_rep)
                    safe_set(12, 11, ut_day_j + ut_night_j)
                    safe_set(12, 12, ut_day_ori + ut_day_rep + ut_night_ori + ut_night_rep)
                    safe_set(12, 14, ut_day_ori + ut_day_rep + ut_night_ori + ut_night_rep)
                    
                    ut_day_adj = get_val(ut, '주간', 'ORI', 'adj_qty') + get_val(ut, '주간', 'REP', 'adj_qty')
                    ut_night_adj = get_val(ut, '야간/휴일', 'ORI', 'adj_qty') + get_val(ut, '야간/휴일', 'REP', 'adj_qty')
                    safe_set(13, 4, ut_day_adj)
                    safe_set(13, 6, ut_day_adj)
                    safe_set(13, 8, ut_night_adj)
                    safe_set(13, 10, ut_night_adj)
                    safe_set(13, 12, ut_day_adj + ut_night_adj)
                    safe_set(13, 14, ut_day_adj + ut_night_adj)
                    
                    pt = agg['PT']['data']
                    pt_day_j = get_val(pt, '주간', 'ORI', 'joints') + get_val(pt, '주간', 'REP', 'joints')
                    pt_day_ori = get_val(pt, '주간', 'ORI', 'qty')
                    pt_day_rep = get_val(pt, '주간', 'REP', 'qty')
                    pt_night_j = get_val(pt, '야간/휴일', 'ORI', 'joints') + get_val(pt, '야간/휴일', 'REP', 'joints')
                    pt_night_ori = get_val(pt, '야간/휴일', 'ORI', 'qty')
                    pt_night_rep = get_val(pt, '야간/휴일', 'REP', 'qty')
                    
                    safe_set(14, 3, pt_day_j)
                    safe_set(14, 4, pt_day_ori + pt_day_rep)
                    safe_set(14, 6, pt_day_ori + pt_day_rep)
                    safe_set(14, 7, pt_night_j)
                    safe_set(14, 8, pt_night_ori + pt_night_rep)
                    safe_set(14, 10, pt_night_ori + pt_night_rep)
                    safe_set(14, 11, pt_day_j + pt_night_j)
                    safe_set(14, 12, pt_day_ori + pt_day_rep + pt_night_ori + pt_night_rep)
                    safe_set(14, 14, pt_day_ori + pt_day_rep + pt_night_ori + pt_night_rep)
                    
                    pt_day_adj = get_val(pt, '주간', 'ORI', 'adj_qty') + get_val(pt, '주간', 'REP', 'adj_qty')
                    pt_night_adj = get_val(pt, '야간/휴일', 'ORI', 'adj_qty') + get_val(pt, '야간/휴일', 'REP', 'adj_qty')
                    safe_set(15, 4, pt_day_adj)
                    safe_set(15, 6, pt_day_adj)
                    safe_set(15, 8, pt_night_adj)
                    safe_set(15, 10, pt_night_adj)
                    safe_set(15, 12, pt_day_adj + pt_night_adj)
                    safe_set(15, 14, pt_day_adj + pt_night_adj)
                    
                    total_film = rt_total['주간_ORI'] + rt_total['주간_REP'] + rt_total['야간_ORI'] + rt_total['야간_REP']
                    safe_set(16, 11, 0)
                    safe_set(16, 12, total_film)
                    
                sheet_names = wb.sheetnames
                if main_agg:
                    sheet_name = '3. 비파괴검사 현황 (열배관)'
                    if sheet_name in sheet_names:
                        write_ndt_sheet_openpyxl(wb[sheet_name], main_agg)
                        log(f"✅ '{sheet_name}' 시트 기입 완료")
                    else:
                        log(f"⚠️ '{sheet_name}' 시트를 찾을 수 없습니다.")
                        
                if mgmt_agg:
                    sheet_name = '3. 비파괴검사 현황 (관리소)'
                    if sheet_name in sheet_names:
                        write_ndt_sheet_openpyxl(wb[sheet_name], mgmt_agg)
                        log(f"✅ '{sheet_name}' 시트 기입 완료")
                    else:
                        log(f"⚠️ '{sheet_name}' 시트를 찾을 수 없습니다.")
                        
                history = {}
                import json, os
                history_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'daily_work_history.json')
                if os.path.exists(history_path):
                    with open(history_path, 'r', encoding='utf-8') as f:
                        history = json.load(f)
                        
                # (Old PAUT block removed in favor of tagged writer below)
                wb.save(save_path)
                wb.close()
                log(f"\n🎉 저장 완료: {save_path}")
                messagebox.showinfo("완료", f"월간 진도보고서 비파괴검사 현황이 업데이트되었습니다.\n{save_path}")
                import os
                os.startfile(os.path.dirname(save_path))
                
            
                # --- 4.5 NDT 결과서 섹션 태그 기반 자동 기입 ---
                try:
                    import sys as _sys
                    import os as _os
                    _src = _os.path.dirname(_os.path.abspath(__file__))
                    if _src not in _sys.path:
                        _sys.path.insert(0, _src)
                    
                    from site_apps.central.src.tagged_ndt_writer import write_all_tagged_sections
                    target_month_str = f"{year}-{month:02d}"
                    
                    # 저장된 파일을 다시 열어 NDT 기입
                    import openpyxl as _opx
                    wb2 = _opx.load_workbook(save_path)
                    ws2 = wb2.worksheets[0]
                    
                    # 태그 기반 NDT 섹션 기입
                    write_all_tagged_sections(ws2, history, target_month_str, log_func=log)
                    
                    wb2.save(save_path)
                    wb2.close()
                    log("✅ 태그 기반 NDT 결과서 전체 기입 완료")
                    
                    log("✅ NDT 결과서 전체 기입 완료")
                    
                    log("✅ NDT 결과서 전체 기입 완료")
                    
                except Exception as ex:
                    log(f"⚠️ NDT 결과서 기입 오류 (무시됨): {ex}")
                    import traceback
                    log(traceback.format_exc())
                    log("✅ 태그 변환 완료")
                except Exception as ex:
                    log(f"⚠️ 태그 변환 중 오류 (무시됨): {ex}")
                    
            except Exception as e:

                log(f"❌ 오류 발생: {e}")
                import traceback
                log(traceback.format_exc())
                messagebox.showerror("오류", f"내보내기 중 오류: {e}")
        
        btn_frame = ttk.Frame(top)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="📊 내보내기 실행", command=do_export, style='Accent.TButton' if 'Accent.TButton' in self.style.theme_names() else 'TButton').pack(side='left', padx=10)
        ttk.Button(btn_frame, text="취소", command=top.destroy).pack(side='left', padx=10)

    def export_all_daily_usage(self):
        """Export all daily usage records to Excel"""
        try:
            if self.daily_usage_df.empty:
                messagebox.showinfo("알림", "기록된 데이터가 없습니다.")
                return
            
            # Export all columns from the tree structure
            selected_cols = list(self.daily_usage_tree['columns'])
            
            filename = f"전체사용기록_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=filename, filetypes=[("Excel files", "*.xlsx")])
            
            if save_path:
                # Ensure we only try to export columns that exist in the dataframe
                valid_cols = [c for c in selected_cols if c in self.daily_usage_df.columns]
                # If dataframe has extra columns not in tree (like raw data), include them? 
                # Better to stick to what the user sees/defines in tree interface + critical data
                # Actually, export_all usually implies backing up the raw dataframe. 
                # But previous logic used tree columns. Let's stick to tree columns if they exist in DF.
                
                # Fallback: if tree columns are not in DF (e.g. calculated columns), we might have issues.
                # The daily_usage_df has raw data. The tree has formatted data.
                # daily_usage_df columns: Date, Site, MaterialID, Usage, Note, etc. + RTK_..., NDT_...
                # Tree columns: 날짜, 현장, 작업자... OT합계...
                
                # If we want "All Daily Usage", we should probably export the RAW dataframe for backup purposes, 
                # OR the formatted view for reporting.
                # The previous code seemed to try to export based on tree columns but mapped from DF?
                # "export_df = self.daily_usage_df[selected_cols].copy()"
                # This implies 'selected_cols' MUST exist in daily_usage_df.
                # Let's check daily_usage_df columns again.
                # It has 'Date', 'Site', 'MaterialID'...
                # It does NOT have '날짜', '현장' (Korean names).
                # So the previous code 'self.daily_usage_df[selected_cols]' would have FAILED if selected_cols came from tree columns!
                
                # WAIT. The user said "remove column selection". 
                # The previous working code (before my changes) probably did:
                # export_df = self.daily_usage_df.copy()
                # Let's revert to a safe "Backup" style export for "All Data".
                
                export_df = self.daily_usage_df.copy()
                export_df = self.clean_df_export(export_df)
                self.save_df_to_excel_autofit(export_df, save_path, "전체사용기록")
                
                export_df = self.clean_df_export(export_df)
                self.save_df_to_excel_autofit(export_df, save_path, "전체기록")
                messagebox.showinfo("완료", f"전체 기록이 엑셀로 저장되었습니다.\n{save_path}")
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 내보내기 중 오류가 발생했습니다: {e}")

    def export_invoice_excel(self):
        """Export current filtered daily usage to official Invoice Excel (가산가평기성양식)"""
        try:
            import win32com.client as win32
        except ImportError:
            messagebox.showerror("오류", "pywin32 라이브러리가 설치되어 있지 않습니다.\n(pip install pywin32)")
            return

        selected_items = self.daily_usage_tree.selection()
        if not selected_items:
            items_to_export = self.daily_usage_tree.get_children()
        else:
            items_to_export = selected_items

        if not items_to_export:
            messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
            return

        records = []
        cols = list(self.daily_usage_tree['columns'])
        
        def safe_float(val):
            try: return float(str(val).replace(',', '').strip())
            except: return 0.0
            
        def safe_int(val):
            try: return int(float(str(val).replace(',', '').strip()))
            except: return 0

        for item in items_to_export:
            tags = self.daily_usage_tree.item(item, 'tags')
            if tags and 'total' in tags: continue
            
            vals = self.daily_usage_tree.item(item, 'values')
            if not vals: continue
            
            row = dict(zip(cols, vals))
            
            records.append({
                "date": row.get("날짜", ""),
                "loc": f"{row.get('현장', '')} / {row.get('검사품명', '')}".strip(" /"),
                "ndt_type": row.get("검사방법", ""),
                "work_time": row.get("작업형태", "일반") or "일반",
                "material_type": row.get("품목명", ""),
                "qty": safe_float(row.get("수량", 0)),
                "unit": row.get("단위", ""),
                "corr": safe_float(row.get("보정계수", 1)),
                "adjusted_qty": safe_float(row.get("환산물량", 0)),
                "mat_cost": safe_int(row.get("재료비", 0)),
                "lab_cost": safe_int(row.get("인건비", 0)),
                "overhead": safe_int(row.get("제경비", 0)),
                "tech": safe_int(row.get("기술료", 0)),
                "subtotal": safe_int(row.get("검사비", 0))
            })

        if not records:
            messagebox.showwarning("기록 없음", "내보낼 작업 기록이 없습니다.")
            return

        top = tk.Toplevel(self.root)
        top.title("실비 정산 및 기성청구서 출력")
        top.geometry("350x400")
        top.transient(self.root)
        top.grab_set()

        ttk.Label(top, text="엑셀 청구서 하단에 합산될\n추가 실비정산 금액을 입력하세요.\n(세액 미포함 금액)", justify=tk.CENTER).pack(pady=15)

        exp_vars = {}
        for label in ["장비손료", "안전관리비", "주재비 및 출장여비", "도서인쇄비"]:
            f = ttk.Frame(top)
            f.pack(fill=tk.X, padx=20, pady=5)
            ttk.Label(f, text=label, width=15).pack(side=tk.LEFT)
            var = tk.IntVar(value=0)
            ttk.Entry(f, textvariable=var, justify='right').pack(side=tk.RIGHT, expand=True, fill=tk.X)
            exp_vars[label] = var

        def do_export():
            extra_expenses = {k: v.get() for k, v in exp_vars.items()}
            top.destroy()
            
            default_name = f"기성청구내역서_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_name, filetypes=[("Excel File", "*.xlsx")], title="정식 기성청구 엑셀 양식으로 저장")
            if not filepath: return

            try:
                excel = win32.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                wb = excel.Workbooks.Add()
                ws = wb.ActiveSheet
                ws.Name = "기성청구내역서"

                ws.Range("A1:O2").Merge()
                ws.Range("A1").Value = "비파괴검사기술용역 기성청구 내역서"
                ws.Range("A1").Font.Size = 20
                ws.Range("A1").Font.Bold = True
                ws.Range("A1").HorizontalAlignment = -4108
                ws.Range("A1").VerticalAlignment = -4108

                ws.Range("A4:B4").Merge()
                ws.Range("A4").Value = "공 사 명 :"
                ws.Range("A4").Font.Bold = True
                ws.Range("C4:J4").Merge()
                project_name = records[0]["loc"].split('/')[0].strip() if records else ""
                if not project_name: project_name = "가산~가평 천연가스 공급시설 건설공사 비파괴검사 기술용역"
                ws.Range("C4").Value = project_name

                ws.Range("L4:M4").Merge()
                ws.Range("L4").Value = "청구일자 :"
                ws.Range("L4").Font.Bold = True
                ws.Range("L4").HorizontalAlignment = -4152
                ws.Range("N4:O4").Merge()
                ws.Range("N4").Value = datetime.datetime.now().strftime('%Y년 %m월 %d일')

                headers = ["No.", "검사일자", "작업구간", "검사종류", "규격/자재", "근무형태", "실물량", "단위", "보정계수", "환산물량", 
                           "재료비", "직접인건비", "제경비", "기술료", "공급가액소계"]
                
                start_row = 6
                for col, h in enumerate(headers, start=1):
                    cell = ws.Cells(start_row, col)
                    cell.Value = h
                    cell.Font.Bold = True
                    cell.Interior.Color = 14277081
                    cell.HorizontalAlignment = -4108
                    cell.Borders.LineStyle = 1

                widths = [4, 11, 20, 9, 22, 9, 8, 5, 9, 9, 11, 12, 11, 11, 13]
                for idx, w in enumerate(widths, start=1):
                    ws.Columns(idx).ColumnWidth = w

                current_row = start_row + 1
                total_mat = total_lab = total_ovr = total_tech = total_sub = 0
                item_idx = 1

                for g_type in ["RT", "UT", "PT", "MT", "ETC"]:
                    if g_type == "ETC":
                        group_records = [r for r in records if not any(r["ndt_type"].upper().startswith(t) for t in ["RT", "UT", "PT", "MT"])]
                    else:
                        group_records = [r for r in records if r["ndt_type"].upper().startswith(g_type)]
                        
                    if not group_records: continue
                    
                    sub_mat = sub_lab = sub_ovr = sub_tech = sub_sub = 0
                    for r in group_records:
                        ws.Cells(current_row, 1).Value = item_idx
                        ws.Cells(current_row, 2).Value = r["date"]
                        ws.Cells(current_row, 3).Value = r["loc"].split()[0] if isinstance(r["loc"], str) and r["loc"] else r["loc"]
                        ws.Cells(current_row, 4).Value = r["ndt_type"]
                        ws.Cells(current_row, 5).Value = r["material_type"]
                        ws.Cells(current_row, 6).Value = r["work_time"]
                        ws.Cells(current_row, 7).Value = r["qty"]
                        ws.Cells(current_row, 8).Value = r["unit"]
                        ws.Cells(current_row, 9).Value = r["corr"]
                        ws.Cells(current_row, 10).Value = r["adjusted_qty"]
                        ws.Cells(current_row, 11).Value = r["mat_cost"]
                        ws.Cells(current_row, 12).Value = r["lab_cost"]
                        ws.Cells(current_row, 13).Value = r["overhead"]
                        ws.Cells(current_row, 14).Value = r["tech"]
                        ws.Cells(current_row, 15).Value = r["subtotal"]
                        
                        for c in range(1, 16):
                            cell = ws.Cells(current_row, c)
                            cell.Borders.LineStyle = 1
                            if c <= 4 or c == 6 or c == 8: cell.HorizontalAlignment = -4108
                            elif c == 5 or c == 3: cell.HorizontalAlignment = -4131
                            else: cell.NumberFormat = "#,##0" if c >= 11 else "0.00"
                        
                        sub_mat += r["mat_cost"]; sub_lab += r["lab_cost"]; sub_ovr += r["overhead"]
                        sub_tech += r["tech"]; sub_sub += r["subtotal"]
                        item_idx += 1; current_row += 1
                    
                    ws.Range(ws.Cells(current_row, 1), ws.Cells(current_row, 10)).Merge()
                    ws.Cells(current_row, 1).Value = f"[{g_type}] 검사 소계"
                    ws.Cells(current_row, 1).HorizontalAlignment = -4108
                    ws.Cells(current_row, 1).Font.Bold = True
                    
                    ws.Cells(current_row, 11).Value = sub_mat
                    ws.Cells(current_row, 12).Value = sub_lab
                    ws.Cells(current_row, 13).Value = sub_ovr
                    ws.Cells(current_row, 14).Value = sub_tech
                    ws.Cells(current_row, 15).Value = sub_sub
                    
                    for c in range(1, 16):
                        cell = ws.Cells(current_row, c)
                        cell.Borders.LineStyle = 1
                        cell.Font.Bold = True
                        cell.Interior.Color = 15987699
                        if c >= 11: cell.NumberFormat = "#,##0"
                    
                    total_mat += sub_mat; total_lab += sub_lab; total_ovr += sub_ovr
                    total_tech += sub_tech; total_sub += sub_sub
                    current_row += 1

                ws.Range(ws.Cells(current_row, 1), ws.Cells(current_row, 10)).Merge()
                ws.Cells(current_row, 1).Value = "검사 비용 합계"
                ws.Cells(current_row, 1).HorizontalAlignment = -4108
                ws.Cells(current_row, 1).Font.Bold = True
                
                ws.Cells(current_row, 11).Value = total_mat
                ws.Cells(current_row, 12).Value = total_lab
                ws.Cells(current_row, 13).Value = total_ovr
                ws.Cells(current_row, 14).Value = total_tech
                ws.Cells(current_row, 15).Value = total_sub
                
                for c in range(1, 16):
                    cell = ws.Cells(current_row, c)
                    cell.Borders.LineStyle = 1
                    cell.Font.Bold = True
                    cell.Interior.Color = 14277081
                    if c >= 11: cell.NumberFormat = "#,##0"
                        
                current_row += 1
                total_extra = 0
                for name, val in extra_expenses.items():
                    if val > 0:
                        ws.Range(ws.Cells(current_row, 1), ws.Cells(current_row, 14)).Merge()
                        ws.Cells(current_row, 1).Value = f"+ {name}"
                        ws.Cells(current_row, 1).HorizontalAlignment = -4152
                        ws.Cells(current_row, 15).Value = val
                        ws.Cells(current_row, 15).NumberFormat = "#,##0"
                        for c in range(1, 16): ws.Cells(current_row, c).Borders.LineStyle = 1
                        total_extra += val
                        current_row += 1

                grand_subtotal = total_sub + total_extra
                
                ws.Range(ws.Cells(current_row, 1), ws.Cells(current_row, 14)).Merge()
                ws.Cells(current_row, 1).Value = "공급가액 총액"
                ws.Cells(current_row, 1).HorizontalAlignment = -4152
                ws.Cells(current_row, 1).Font.Bold = True
                ws.Cells(current_row, 15).Value = grand_subtotal
                ws.Cells(current_row, 15).NumberFormat = "#,##0"
                ws.Cells(current_row, 15).Font.Bold = True
                for c in range(1, 16): ws.Cells(current_row, c).Borders.LineStyle = 1
                
                current_row += 1
                ws.Range(ws.Cells(current_row, 1), ws.Cells(current_row, 14)).Merge()
                ws.Cells(current_row, 1).Value = "+ 부가가치세 (10%)"
                ws.Cells(current_row, 1).HorizontalAlignment = -4152
                vat_val = int(grand_subtotal * 0.1)
                ws.Cells(current_row, 15).Value = vat_val
                ws.Cells(current_row, 15).NumberFormat = "#,##0"
                for c in range(1, 16): ws.Cells(current_row, c).Borders.LineStyle = 1
                
                current_row += 1
                ws.Range(ws.Cells(current_row, 1), ws.Cells(current_row, 14)).Merge()
                ws.Cells(current_row, 1).Value = "최 종 기 성 청 구 액"
                ws.Cells(current_row, 1).HorizontalAlignment = -4152
                ws.Cells(current_row, 1).Font.Bold = True
                ws.Cells(current_row, 1).Font.Size = 12
                
                total_final = grand_subtotal + vat_val
                ws.Cells(current_row, 15).Value = total_final
                ws.Cells(current_row, 15).NumberFormat = "#,##0"
                ws.Cells(current_row, 15).Font.Bold = True
                ws.Cells(current_row, 15).Font.Size = 12
                
                for c in range(1, 16):
                    cell = ws.Cells(current_row, c)
                    cell.Borders.LineStyle = 1
                    cell.Interior.Color = 13434879
                    
                filepath = filepath.replace("/", "\\")
                wb.SaveAs(filepath)
                wb.Close()
                excel.Quit()
                
                messagebox.showinfo("저장 완료", f"엑셀 기성청구 내역서가 성공적으로 생성되었습니다.\n{filepath}")
                os.startfile(filepath)
                
            except Exception as e:
                messagebox.showerror("저장 오류", f"엑셀 파일 생성 중 오류가 발생했습니다.\n{str(e)}")
                try: excel.Quit()
                except: pass

        ttk.Button(top, text="엑셀 출력 실행", command=do_export).pack(pady=20, ipady=5, fill=tk.X, padx=20)



    def setup_keyboard_shortcuts(self):
        """Setup keyboard shortcuts for navigation"""
        # Ctrl+Tab to switch between notebook tabs (forward)
        self.root.bind('<Control-Tab>', self.next_tab)
        # Ctrl+Shift+Tab to switch between notebook tabs (backward)
        self.root.bind('<Control-Shift-Tab>', self.prev_tab)
        
        # Alt+숫자 for direct tab access
        self.root.bind('<Alt-Key-1>', lambda e: self.notebook.select(0))
        self.root.bind('<Alt-Key-2>', lambda e: self.notebook.select(1))
        self.root.bind('<Alt-Key-3>', lambda e: self.notebook.select(2))
        self.root.bind('<Alt-Key-4>', lambda e: self.notebook.select(3))
        self.root.bind('<Alt-Key-5>', lambda e: self.notebook.select(4))
        self.root.bind('<Alt-Key-6>', lambda e: self.notebook.select(5))
        
        # Right-click on notebook for tab reordering
        # Tab interactions removed here, kept in create_widgets
    
    def next_tab(self, event=None):
        """Switch to next tab"""
        current = self.notebook.index(self.notebook.select())
        total = self.notebook.index('end')
        next_tab = (current + 1) % total
        self.notebook.select(next_tab)
        return 'break'
    
    def prev_tab(self, event=None):
        """Switch to previous tab"""
        current = self.notebook.index(self.notebook.select())
        total = self.notebook.index('end')
        prev_tab = (current - 1) % total
        self.notebook.select(prev_tab)
        return 'break'
    
    def show_tab_context_menu(self, event):
        """Show context menu for tab reordering"""
        # Identify which tab was clicked
        try:
            clicked_tab = self.notebook.index(f"@{event.x},{event.y}")
            self.notebook.select(clicked_tab)
            
            # Create context menu
            context_menu = tk.Menu(self.root, tearoff=0)
            
            # Only show "Move Left" if not the first tab
            if clicked_tab > 0:
                context_menu.add_command(label="← 탭 왼쪽으로 이동", 
                                        command=lambda: self.move_tab_left(clicked_tab))
            
            # Only show "Move Right" if not the last tab
            if clicked_tab < self.notebook.index('end') - 1:
                context_menu.add_command(label="탭 오른쪽으로 이동 →", 
                                        command=lambda: self.move_tab_right(clicked_tab))
            
            # Show menu at cursor position
            context_menu.post(event.x_root, event.y_root)
        except:
            pass
    
    def move_tab_left(self, tab_index):
        """Move tab one position to the left"""
        if tab_index > 0:
            # Get tab info
            tab = self.notebook.tabs()[tab_index]
            text = self.notebook.tab(tab_index, "text")
            
            # Remove and reinsert at new position
            self.notebook.insert(tab_index - 1, tab, text=text)
            self.notebook.select(tab_index - 1)
            
            # Save configuration immediately
            self.save_tab_config()
    
    def move_tab_right(self, tab_index):
        """Move tab one position to the right"""
        if tab_index < self.notebook.index('end') - 1:
            # Get tab info
            tab = self.notebook.tabs()[tab_index]
            text = self.notebook.tab(tab_index, "text")
            
            # Remove and reinsert at new position
            self.notebook.insert(tab_index + 2, tab, text=text)
            self.notebook.select(tab_index + 1)
            
            # Save configuration immediately
            self.save_tab_config()

    def on_tab_drag_start(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import on_tab_drag_start_impl
        return on_tab_drag_start_impl(self, *args, **kwargs)

    def on_tab_drag(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import on_tab_drag_impl
        return on_tab_drag_impl(self, *args, **kwargs)

    def normalize_site_name(self, *args, **kwargs):
        from site_apps.central.src.utils.helpers import normalize_site_name_impl
        return normalize_site_name_impl(self, *args, **kwargs)

    def on_tab_drag_end(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import on_tab_drag_end_impl
        return on_tab_drag_end_impl(self, *args, **kwargs)
    
    def on_tab_changed(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import on_tab_changed_impl
        return on_tab_changed_impl(self, *args, **kwargs)
    
    def auto_save_to_list(self, event, combobox, data_list, config_key):
        """Helper to auto-save new entry from combobox to a list and update all related UI"""
        new_val = combobox.get().strip()
        if not new_val:
            return
            
        if new_val not in data_list:
            data_list.append(new_val)
            data_list.sort()
            
            # Trigger app-wide update
            self.refresh_ui_for_list_change(config_key)
            self.save_tab_config()


    def auto_save_worktime(self, event, entry, config_key):
        """Helper to auto-save worktime values and support copy functionality"""
        worktime_value = entry.get().strip()
        if not worktime_value:
            return
            
        # Initialize worktimes list if not exists
        if not hasattr(self, 'worktimes'):
            self.worktimes = []
            
        # Check if this value already exists
        if worktime_value not in self.worktimes:
            self.worktimes.append(worktime_value)
            self.worktimes.sort()
            
            # Save to config and update UI
            try:
                self.refresh_ui_for_list_change('worktimes')
                self.save_tab_config()
            except Exception as e:
                print(f"Failed to save worktime config: {e}")
        
        # Calculate and update corresponding OT field
        self.calculate_and_update_ot(worktime_value, entry)
        
        # Add copy functionality - select all text on focus and bind Ctrl+C (only once)
        entry.select_range(0, tk.END)
        if not hasattr(entry, '_copy_bound'):
            entry.bind('<Control-c>', lambda e: self.copy_worktime(entry))
            entry._copy_bound = True
        
        # Store current value for copy functionality
        entry.last_value = worktime_value

    def calculate_and_update_ot(self, worktime_value, worktime_entry):
        """Calculate OT amount based on worktime and update corresponding OT field"""
        try:
            current_date = self.ent_daily_date.get_date()
            ot_hours, amount = self._calculate_ot_from_worktime(worktime_value, current_date)
            
            ot_entry = self.get_corresponding_ot_field(worktime_entry)
            if ot_entry:
                if amount > 0:
                    ot_entry.delete(0, tk.END)
                    ot_entry.insert(0, f"{amount:,}")
                else:
                    ot_entry.delete(0, tk.END)
                    # If it was cleared, but hours > 0 (very small amount?), maybe set to 0
                    if ot_hours > 0: ot_entry.insert(0, "0")
        except Exception as e:
            print(f"Error calculating OT: {e}")

    def get_corresponding_ot_field(self, worktime_entry):
        """Get corresponding OT field for a worktime entry"""
        # Map worktime entries to OT entries
        worktime_to_ot = {
            'ent_worktime1': 'ent_ot1', 'ent_worktime2': 'ent_ot2', 
            'ent_worktime3': 'ent_ot3', 'ent_worktime4': 'ent_ot4',
            'ent_worktime5': 'ent_ot5', 'ent_worktime6': 'ent_ot6',
            'ent_worktime7': 'ent_ot7', 'ent_worktime8': 'ent_ot8',
            'ent_worktime9': 'ent_ot9', 'ent_worktime10': 'ent_ot10'
        }
            
        # Find worktime entry name by checking object IDs
        worktime_id = id(worktime_entry)
            
        worktime_attrs = [f'ent_worktime{i}' for i in range(1, 11)]
        for attr_name in worktime_attrs:
            if hasattr(self, attr_name):
                attr_value = getattr(self, attr_name)
                if id(attr_value) == worktime_id:
                    ot_attr_name = worktime_to_ot.get(attr_name)
                    if ot_attr_name and hasattr(self, ot_attr_name):
                        return getattr(self, ot_attr_name)
                    break
            
        return None

    def copy_worktime(self, entry):
        """Copy worktime value to clipboard"""
        try:
            worktime_value = entry.get().strip()
            if worktime_value:
                self.root.clipboard_clear()
                self.root.clipboard_append(worktime_value)
                # Show brief feedback
                entry.delete(0, tk.END)
                entry.insert(0, worktime_value)
        except Exception as e:
            pass  # Silently handle clipboard errors

    def auto_save_ot(self, event, entry, config_key):
        """Helper to auto-save OT values and support copy functionality"""
        ot_value = entry.get().strip()
        if not ot_value:
            return
                
        # Initialize ot_times list if not exists
        if not hasattr(self, 'ot_times'):
            self.ot_times = []
            
        # Calculate OT amount for any input
        calculated_amount = self.calculate_ot_amount(ot_value)
        if calculated_amount:
            # Update entry with calculated amount
            entry.delete(0, tk.END)
            display_value = f"{ot_value} ({calculated_amount:,}원)"
            entry.insert(0, display_value)
            ot_value = display_value
            print(f"OT calculated: {ot_value}")  # Debug print
            
        # Check if this value already exists
        if ot_value not in self.ot_times:
            self.ot_times.append(ot_value)
            self.ot_times.sort()
            
        # Save to config with error handling - only save when list changes
        try:
            self.save_tab_config()
        except Exception as e:
            print(f"Failed to save OT config: {e}")
            
        # Add copy functionality - select all text on focus and bind Ctrl+C (only once)
        entry.select_range(0, tk.END)
        if not hasattr(entry, '_copy_bound'):
            entry.bind('<Control-c>', lambda e: self.copy_ot(entry))
            entry._copy_bound = True
            
        # Store current value for copy functionality
        entry.last_value = ot_value

    def calculate_ot_amount(self, *args, **kwargs):
        from site_apps.central.src.models.worker_model import calculate_ot_amount_impl
        return calculate_ot_amount_impl(self, *args, **kwargs)

    def _parse_ot_hours(self, *args, **kwargs):
        from site_apps.central.src.models.worker_model import _parse_ot_hours_impl
        return _parse_ot_hours_impl(self, *args, **kwargs)

    def _calculate_split_ot_hours(self, *args, **kwargs):
        from site_apps.central.src.models.worker_model import _calculate_split_ot_hours_impl
        return _calculate_split_ot_hours_impl(self, *args, **kwargs)

    def copy_ot(self, entry):
        """Copy OT value to clipboard"""
        try:
            ot_value = entry.get().strip()
            if ot_value:
                self.root.clipboard_clear()
                self.root.clipboard_append(ot_value)
                # Show brief feedback
                entry.delete(0, tk.END)
                entry.insert(0, ot_value)
        except Exception as e:
            pass  # Silently handle clipboard errors


    def save_tab_config(self, force=False):
        """Save current tab configuration to memory and disk"""
        try:
            # [STABILITY] Skip saving if app is not fully ready (prevents overwriting saved tab index during init)
            if not getattr(self, 'is_ready', False) and not force:
                return

            # Force update to ensure all coordinates and sizes are accurate
            try:
                self.root.update_idletasks()
            except:
                pass

            # Initialize or keep existing tab_config
            if not hasattr(self, 'tab_config'):
                self.tab_config = {}
                
            # Update with current core values
            current_tab_idx = 0
            current_tab_text = ""
            try:
                current_tab_widget = self.notebook.select()
                current_tab_idx = self.notebook.index(current_tab_widget)
                current_tab_text = self.notebook.tab(current_tab_widget, "text")
            except:
                current_tab_idx = self.tab_config.get('selected_tab', 0)
                current_tab_text = self.tab_config.get('selected_tab_text', "")

            # Get current tab order
            tab_order = []
            for tab_id in self.notebook.tabs():
                tab_order.append(self.notebook.tab(tab_id, "text"))

            daily_usage_sash_pos = None
            if hasattr(self, 'daily_usage_paned'):
                try: daily_usage_sash_pos = 500
                except: pass

            daily_history_sash_pos = None
            if hasattr(self, 'daily_history_paned'):
                try: daily_history_sash_pos = self.daily_history_paned.sashpos(0)
                except: pass

            bottom_dashboard_sashes = []
            if hasattr(self, 'bottom_dashboard') and isinstance(self.bottom_dashboard, ttk.PanedWindow):
                try:
                    bottom_dashboard_sashes = [self.bottom_dashboard.sashpos(0), self.bottom_dashboard.sashpos(1)]
                except: pass

            entry_inner_frame_height = None
            if hasattr(self, 'entry_inner_frame'):
                try: entry_inner_frame_height = self.entry_inner_frame.winfo_height()
                except: pass
                
            window_state = 'normal'
            window_width = 1200
            window_height = 800
            try:
                window_state = self.root.state()
                window_width = self.root.winfo_width()
                window_height = self.root.winfo_height()
            except: pass

            self.tab_config.update({
                'selected_tab': current_tab_idx,
                'selected_tab_text': current_tab_text,
                'tab_order': tab_order,
                'sites': self.sites,
                'daily_units': self.daily_units,
                'users': getattr(self, 'users', []),
                'companies': getattr(self, 'companies', []),
                'warehouses': getattr(self, 'warehouses', []),
                'equipments': getattr(self, 'equipments', []),
                'test_items': getattr(self, 'test_items', []),
                'applied_codes': getattr(self, 'applied_codes', []),
                'vehicles': getattr(self, 'vehicles', []),
                'materials': getattr(self, 'carestream_films', []),
                'worktimes': getattr(self, 'worktimes', []),
                'ot_times': getattr(self, 'ot_times', []),
                'layout_locked': getattr(self, 'layout_locked', False),
                'resolution_locked': getattr(self, 'resolution_locked', False),
                'locked_width': getattr(self, 'locked_width', 1200),
                'locked_height': getattr(self, 'locked_height', 800),
                'daily_usage_sash_locked': getattr(self, 'daily_usage_sash_locked', False),
                'daily_usage_sash_pos': daily_usage_sash_pos,
                'daily_history_sash_pos': daily_history_sash_pos,
                'bottom_dashboard_sashes': bottom_dashboard_sashes,
                'entry_inner_frame_height': entry_inner_frame_height,
                'history_visible_cols': getattr(self, 'manual_visible_cols', []),
                'monthly_visible_cols': getattr(self, 'monthly_visible_cols', []),
                'budget_view_visible_cols': getattr(self, 'budget_view_visible_cols', []),
                'budget_view_heading_aliases': getattr(self, 'budget_view_heading_aliases', {}),
                'budget_view_custom_columns': getattr(self, 'budget_view_custom_columns', []),
                'window_state': window_state,
                'window_width': window_width,
                'window_height': window_height
            })
            
            # Save current stock column widths
            self.tab_config['stock_col_widths'] = {}
            if hasattr(self, 'stock_tree'):
                for col in self.stock_tree['columns']:
                    self.tab_config['stock_col_widths'][col] = self.stock_tree.column(col, 'width')

            # Save in/out history column widths
            self.tab_config['inout_col_widths'] = {}
            if hasattr(self, 'inout_tree'):
                for col in self.inout_tree['columns']:
                    self.tab_config['inout_col_widths'][col] = self.inout_tree.column(col, 'width')

            # Save recent entries column widths
            self.tab_config['tv_recent_col_widths'] = {}
            if hasattr(self, 'tv_recent'):
                for col in self.tv_recent['columns']:
                    self.tab_config['tv_recent_col_widths'][col] = self.tv_recent.column(col, 'width')

            # Save daily usage column widths
            self.tab_config['daily_usage_col_widths'] = {}
            if hasattr(self, 'daily_usage_tree'):
                for col in self.daily_usage_tree['columns']:
                    self.tab_config['daily_usage_col_widths'][col] = self.daily_usage_tree.column(col, 'width')
            
            # Save monthly usage column widths
            self.tab_config['monthly_usage_col_widths'] = {}
            if hasattr(self, 'monthly_usage_tree'):
                for col in self.monthly_usage_tree['columns']:
                    self.tab_config['monthly_usage_col_widths'][col] = self.monthly_usage_tree.column(col, 'width')
            
            # [NEW] Save summary table column widths
            self.tab_config['site_summary_col_widths'] = {}
            if hasattr(self, 'site_summary_tree'):
                for col in self.site_summary_tree['columns']:
                    self.tab_config['site_summary_col_widths'][col] = self.site_summary_tree.column(col, 'width')
            
            self.tab_config['worker_summary_col_widths'] = {}
            if hasattr(self, 'worker_summary_tree'):
                for col in self.worker_summary_tree['columns']:
                    self.tab_config['worker_summary_col_widths'][col] = self.worker_summary_tree.column(col, 'width')
            
            # [NEW] Save report column widths
            self.tab_config['report_col_widths'] = {}
            if hasattr(self, 'report_tree'):
                for col in self.report_tree['columns']:
                    self.tab_config['report_col_widths'][col] = self.report_tree.column(col, 'width')

            # Save budget site performance column widths
            self.tab_config['budget_view_col_widths'] = {}
            if hasattr(self, 'budget_view_tree'):
                for col in self.budget_view_tree['columns']:
                    self.tab_config['budget_view_col_widths'][col] = self.budget_view_tree.column(col, 'width')
            
            # Save tab order
            self.tab_config['tab_order'] = []
            for tab in self.notebook.tabs():
                self.tab_config['tab_order'].append(self.notebook.tab(tab, "text"))
                
            # [NEW] Save equipment list
            self.tab_config['equipments'] = self.equipments
            
            # Save Draggable items
            # Prepare in-memory geometries for merging (capture current states)
            current_geometries = {}
            
            for key, widget in self.draggable_items.items():
                try:
                    # [LAYOUT FIX] 핵심 박스는 좌표 저장 제외
                    # 재시작 시 기본 grid 배치로 복원하여 겹침/잘림 방지
                    if key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
                        continue

                    manager = widget.winfo_manager()
                    # Capture everything that is visible and managed
                    if manager in ['place', 'grid']:
                        current_geometries[key] = {
                            'x': widget.winfo_x(),
                            'y': widget.winfo_y(),
                            'width': widget.winfo_width(),
                            'height': widget.winfo_height(),
                            'hidden': False
                        }
                        
                        if hasattr(widget, '_label_widget'):
                            current_geometries[key]['custom_label'] = widget._label_widget.cget('text')
                        if hasattr(widget, '_manage_list_key') and widget._manage_list_key:
                            current_geometries[key]['manage_list_key'] = widget._manage_list_key
                        if key.startswith('clone_'):
                            current_geometries[key]['is_clone'] = True
                            current_geometries[key]['widget_class_name'] = widget._widget_class.__name__
                            saved_kwargs = widget._widget_kwargs.copy()
                            if 'values' in saved_kwargs: del saved_kwargs['values']
                            # Remove un-serializable callable objects (like on_save callbacks)
                            keys_to_del = [k for k, v in saved_kwargs.items() if callable(v)]
                            for k in keys_to_del: del saved_kwargs[k]
                            current_geometries[key]['widget_kwargs'] = saved_kwargs
                        if key.startswith('memo_'):
                            current_geometries[key]['text'] = self.memos[key]['text_widget'].get('1.0', 'end-1c')
                            current_geometries[key]['memo_title'] = self.memos[key]['title_entry'].get()
                        if key in getattr(self, 'checklists', {}):
                            current_geometries[key]['checklist_title'] = self.checklists[key]['title_entry'].get()
                            items_data = []
                            for child in self.checklists[key]['item_frame'].winfo_children():
                                if hasattr(child, '_checklist_data'):
                                    items_data.append({
                                        'text': child._checklist_data['entry'].get(),
                                        'checked': child._checklist_data['var'].get()
                                    })
                            current_geometries[key]['checklist_items'] = items_data
                        if key in getattr(self, 'vehicle_inspections', {}):
                            current_geometries[key]['vehicle_data'] = self.vehicle_inspections[key].get_data()
                    elif manager == '': # Hidden or not managed
                        # Only mark as hidden if it was previously in place/grid (to avoid junk)
                        if key in self.tab_config.get('draggable_geometries', {}):
                            current_geometries[key] = {'hidden': True}
                except Exception as e:
                    print(f"Skipping save for draggable widget {key}: {e}")

            # [DEFINITIVE FIX] Deep merge current_geometries into disk config
            final_config = {}
            if os.path.exists(self.config_path):
                try:
                    with open(self.config_path, 'r', encoding='utf-8') as f:
                        final_config = json.load(f)
                except: pass
            
            # Use disk's geometries as base, then update with WHAT WE JUST CAPTURED in memory
            disk_geos = final_config.get('draggable_geometries', {})

            # [LAYOUT FIX] 과거에 저장된 핵심 박스 좌표도 제거
            for core_key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
                if core_key in disk_geos:
                    del disk_geos[core_key]

            disk_geos.update(current_geometries)
            
            # Apply other tab state updates
            final_config.update(self.tab_config)
            final_config['draggable_geometries'] = disk_geos
            # [NEW] Save hidden_sites to config
            final_config['hidden_sites'] = list(getattr(self, 'hidden_sites', []))
            
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(final_config, f, ensure_ascii=False, indent=2)
            
            # Update in-memory copy as well
            self.tab_config = final_config
            
            print(f"Configuration saved. Locked: {self.tab_config.get('layout_locked')}, Tab: {self.tab_config.get('selected_tab')}")
        except Exception as e:
            print(f"Failed to save tab config: {e}")

    
    def load_tab_config(self):
        """Load and restore tab configuration"""
        try:
            if not hasattr(self, 'tab_config'):
                self.tab_config = {}
                
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    self.tab_config = json.load(f)
                    print(f"DEBUG: Loaded config raw: {self.tab_config}")
                
                config = self.tab_config
                
                # Restore lists - Respecting user deletions
                if 'daily_units' in config:
                    loaded = config['daily_units']
                    if loaded: # If user has a saved list (even if they deleted items)
                        self.daily_units = list(dict.fromkeys(loaded))
                    else: # Fallback only if the list is literally empty
                        self.daily_units = ['EA', 'CAN', 'SET', 'KG', 'M', '매', 'I/D', 'P,M,I/D', 'M,I/D', 'Point', 'Meter', 'Inch', 'Dia']
                
                # Apply high-level lock states immediately to internal variables
                self.layout_locked = config.get('layout_locked', False)
                self.resolution_locked = config.get('resolution_locked', False)
                self.daily_usage_sash_locked = config.get('daily_usage_sash_locked', False)
                self.locked_width = config.get('locked_width', 1200)
                self.locked_height = config.get('locked_height', 800)

                # Get current tab order
                current_order = []
                for tab in self.notebook.tabs():
                    tab_index = self.notebook.index(tab)
                    tab_text = self.notebook.tab(tab_index, "text")
                    current_order.append((tab_text, tab))
                
                # Create mapping from text to tab widget
                tab_map = {text: tab for text, tab in current_order}
                
                # Restore tab order if saved order exists
                saved_order = config.get('tab_order', [])
                if saved_order:
                    # [ROBUST] Reorder tabs according to saved order (relaxed length check)
                    for i, tab_text in enumerate(saved_order):
                        if tab_text in tab_map:
                            tab = tab_map[tab_text]
                            current_pos = self.notebook.index(tab)
                            if current_pos != i:
                                self.notebook.insert(i, tab, text=tab_text)
                
                # Restore selected tab
                # [USER REQUEST] Always start at Daily Usage tab
                selected_idx = 4
                selected_text = "현장별 일일 사용량 기입"
                
                tab_restored = False
                if selected_text:
                    # [ROBUST] Find tab by name instead of index
                    for i in range(self.notebook.index('end')):
                        if self.notebook.tab(i, "text") == selected_text:
                            try:
                                self.notebook.select(i)
                                tab_restored = True
                                break
                            except: pass
                
                if not tab_restored and 0 <= selected_idx < self.notebook.index('end'):
                    try:
                        self.notebook.select(selected_idx)
                    except: pass
                
                # Force update after selection so the tab is rendered and children computed
                self.root.update_idletasks()
                
                # Restore sites list - use in-place update
                self.sites[:] = config.get('sites', [])
                # If sites list is empty, try to populate from current daily_usage_df
                if not self.sites and not self.daily_usage_df.empty:
                    # [FIX] Collect and normalize sites in-place to maintain widget references and resolve crash
                    self.sites[:] = sorted(list(self.daily_usage_df['Site'].dropna().astype(str).apply(self.normalize_site_name).unique()))
                # [NEW] Restore hidden sites list and ensure requested defaults are present
                saved_hidden = config.get('hidden_sites', [])
                for s in ["초안", "롯데현장"]:
                    if s not in saved_hidden: saved_hidden.append(s)
                self.hidden_sites[:] = saved_hidden
                
                # Restore users list - use in-place update to preserve references
                self.users[:] = config.get('users', [])
                
                # [MIGRATION] Load and immediately migrate worktimes to ensure 익일 marker
                raw_worktimes = config.get('worktimes', [])
                self.worktimes = self._migrate_worktimes(raw_worktimes)
                
                # Restore vehicles, equipments, and preferred materials
                raw_vehicles = config.get('vehicles', [])
                updated_vehicles = []
                for v in raw_vehicles:
                    base = v.split('(')[0].strip()
                    if any(t in base for t in ['81두1580', '89보4028', '90너4889']):
                        updated_vehicles.append(f"{base} (탑차)")
                    else:
                        updated_vehicles.append(f"{base} (스타렉스)")
                self.vehicles[:] = sorted(list(set(updated_vehicles)))
                
                self.equipments[:] = config.get('equipments', [])
                self.test_items[:] = config.get('test_items', [])
                self.applied_codes[:] = config.get('applied_codes', [])
                if 'materials' in config:
                    self.carestream_films[:] = config['materials']
                
                # [NEW] Restore companies list
                self.companies[:] = config.get('companies', [])
                if hasattr(self, 'cb_daily_company'):
                    self.cb_daily_company['values'] = sorted(self.companies)

                
                # Refresh all WorkerDataGroup dropdowns with the migrated list
                for i in range(1, 11):
                    group_attr = f'worker_group{i}'
                    if hasattr(self, group_attr):
                        getattr(self, group_attr).update_time_list(self.worktimes)

                self.ot_times = config.get('ot_times', [])

                # [CLEANUP] Clean up users list from shift markers
                import re
                cleaned_users = []
                for user in self.users:
                    if re.match(r"^\((주간|야간|휴일)\)$", user.strip()):
                        continue
                    match = re.match(r"\((주간|야간|휴일)\)\s*(.*)", user.strip())
                    if match:
                        actual_name = match.group(2).strip()
                        if actual_name and actual_name not in cleaned_users:
                            cleaned_users.append(actual_name)
                    elif user.strip() and user.strip() not in cleaned_users:
                        cleaned_users.append(user.strip())
                
                if len(cleaned_users) != len(self.users):
                    self.users[:] = cleaned_users
                    # Don't save yet, let the normal flow handle it
                
                # Restore history visibility if saved
                # [FIX] Mismatch between 'history_visible_cols' and 'daily_usage_visible_cols'
                self.manual_visible_cols = config.get('daily_usage_visible_cols', config.get('history_visible_cols', []))
                
                # [MIGRATION] If 'OT합계' is in saved cols, replace with 'OT시간', 'OT금액'
                if 'OT합계' in self.manual_visible_cols:
                    index = self.manual_visible_cols.index('OT합계')
                    self.manual_visible_cols.pop(index)
                    if 'OT시간' not in self.manual_visible_cols:
                        self.manual_visible_cols.insert(index, 'OT시간')
                    if 'OT금액' not in self.manual_visible_cols:
                        self.manual_visible_cols.insert(index + 1, 'OT금액')
                
                # [MIGRATION] Translate English names to Korean
                translations = {
                    'Date': '날짜', 'Site': '현장', 'User': '작업자', 'WorkTime': '작업시간',
                    'Equipment': '장비명', 'Method': '검사방법', 'Remark': '비고', 
                    'Note': '비고', 'EntryTime': '입력시간', 'MaterialName': '품목명',
                    '제경비율': '제경비', '기술료율': '기술료',
                    '차량번호': '차량번호', '주행거리': '주행거리', '차량점검': '차량점검', '차량비고': '차량비고'
                }
                for i, col in enumerate(self.manual_visible_cols):
                    if col in translations:
                        self.manual_visible_cols[i] = translations[col]
                
                if self.manual_visible_cols and hasattr(self, 'daily_usage_tree'):
                    # Append new NDT columns to saved config if they are missing
                    for nc in ['작업형태', '조건1', '조건2', '보정계수', '제경비', '기술료', '환산물량', '재료비', '인건비']:
                        if nc not in self.manual_visible_cols:
                            self.manual_visible_cols.append(nc)
                            
                    try:
                        self.daily_usage_tree['displaycolumns'] = [c for c in self.manual_visible_cols if c in self.daily_usage_tree['columns']]
                    except: pass
                
                # Restore monthly history visibility if saved
                self.monthly_visible_cols = config.get('monthly_visible_cols', [])
                if self.monthly_visible_cols and hasattr(self, 'monthly_usage_tree'):
                    try:
                        self.monthly_usage_tree['displaycolumns'] = [c for c in self.monthly_visible_cols if c in self.monthly_usage_tree['columns']]
                    except: pass

                # Restore budget site performance view visibility if saved
                self.budget_view_visible_cols = config.get('budget_view_visible_cols', [])
                self.budget_view_heading_aliases = config.get('budget_view_heading_aliases', {})
                self.budget_view_custom_columns = config.get('budget_view_custom_columns', [])
                if hasattr(self, 'budget_view_tree'):
                    try:
                        self._refresh_budget_view_tree_columns(reload_data=False)
                    except:
                        pass
                if self.budget_view_visible_cols and hasattr(self, 'budget_view_tree'):
                    try:
                        visible = [c for c in self.budget_view_visible_cols if c in self.budget_view_tree['columns']]
                        self.budget_view_tree['displaycolumns'] = visible if visible else self.budget_view_tree['columns']
                    except:
                        pass
                
                # Restore warehouses list
                self.warehouses[:] = config.get('warehouses', [])
                if not self.warehouses and not self.materials_df.empty:
                    self.warehouses = sorted(self.materials_df['창고'].dropna().unique().tolist())
                    self.warehouses = [str(w).strip() for w in self.warehouses if str(w).strip()]
                
                self.equipments[:] = config.get('equipments', [])
                if not self.equipments and not self.daily_usage_df.empty and '장비명' in self.daily_usage_df.columns:
                    self.equipments = sorted(self.daily_usage_df['장비명'].dropna().unique().tolist())
                    self.equipments = [str(e).strip() for e in self.equipments if str(e).strip()]
                
                # Vehicle list
                self.vehicles[:] = config.get('vehicles', [])
        
                # 이미지에 제공된 기본 차량 목록 등록
                default_vehicles = [
                    "84저1259", "95가0200", "91주8839", "90서8616",
                    "76마3422", "71고4405", "89보4028", "81두1580",
                    "81루5100", "95가0175", "81도5958", "90너4889"
                ]
                
                # Check based on base name to avoid duplicates
                current_bases = [v.split('(')[0].strip() for v in self.vehicles]
                for dv in default_vehicles:
                    if dv not in current_bases:
                        self.vehicles.append(dv)
                
                # [NEW] Force suffix application on all vehicles before UI update
                final_vehicles = []
                for v in self.vehicles:
                    base = v.split('(')[0].strip()
                    if any(t in base for t in ['81두1580', '89보4028', '90너4889']):
                        final_vehicles.append(f"{base} (탑차)")
                    else:
                        final_vehicles.append(f"{base} (스타렉스)")
                self.vehicles[:] = sorted(list(set(final_vehicles)))
                # [FIX] Update combo box values after loading lists from config
                # Update sites combo boxes
                if hasattr(self, 'cb_trans_site'): 
                    self.cb_trans_site['values'] = sorted(self.sites)
                # Update warehouses combo boxes
                if hasattr(self, 'cb_warehouse'): 
                    self.cb_warehouse['values'] = sorted(self.warehouses)
                # Update users combo boxes
                if hasattr(self, 'ent_user'): 
                    self.ent_user['values'] = sorted(self.users)
                # Update vehicles for VehicleInspectionWidgets
                if hasattr(self, 'fixed_vehicle_widget'):
                    self.fixed_vehicle_widget.update_vehicle_list(sorted(self.vehicles))
                if hasattr(self, 'vehicle_boxes'):
                    for w in self.vehicle_boxes:
                        if hasattr(w, 'update_vehicle_list'):
                            w.update_vehicle_list(sorted(self.vehicles))
                if hasattr(self, 'vehicle_inspections'):
                    for w in self.vehicle_inspections.values():
                        if hasattr(w, 'update_vehicle_list'):
                            w.update_vehicle_list(sorted(self.vehicles))
                
                # Restore stock column widths
                stock_col_widths = config.get('stock_col_widths', {})
                if stock_col_widths and hasattr(self, 'stock_tree'):
                    for col, width in stock_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.stock_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass

                # Restore in/out history column widths
                inout_col_widths = config.get('inout_col_widths', {})
                if inout_col_widths and hasattr(self, 'inout_tree'):
                    for col, width in inout_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.inout_tree.column(col, width=w, minwidth=50, stretch=False)
                        except:
                            pass
                
                # Restore recent entries column widths
                tv_recent_col_widths = config.get('tv_recent_col_widths', {})
                if tv_recent_col_widths and hasattr(self, 'tv_recent'):
                    for col, width in tv_recent_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.tv_recent.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass

                # Restore daily usage column widths
                daily_usage_col_widths = config.get('daily_usage_col_widths', {})
                if daily_usage_col_widths and hasattr(self, 'daily_usage_tree'):
                    for col, width in daily_usage_col_widths.items():
                        try:
                            # Only apply if width is reasonable (e.g. > 10)
                            w = int(width)
                            
                            # All columns are now user-resizable
                            if w > 10:
                                # Enforce minimums for high-precision cols
                                if col == '날짜': w = max(w, 160)
                                elif col == '입력시간': w = max(w, 300)
                                self.daily_usage_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass

                # Restore monthly usage column widths
                monthly_usage_col_widths = config.get('monthly_usage_col_widths', {})
                if monthly_usage_col_widths and hasattr(self, 'monthly_usage_tree'):
                    for col, width in monthly_usage_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.monthly_usage_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass
                
                # [NEW] Restore site summary column widths
                site_summary_col_widths = config.get('site_summary_col_widths', {})
                if site_summary_col_widths and hasattr(self, 'site_summary_tree'):
                    for col, width in site_summary_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.site_summary_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass

                # Restore budget site performance column widths
                budget_view_col_widths = config.get('budget_view_col_widths', {})
                if budget_view_col_widths and hasattr(self, 'budget_view_tree'):
                    for col, width in budget_view_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.budget_view_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass

                # [NEW] Restore worker summary column widths
                worker_summary_col_widths = config.get('worker_summary_col_widths', {})
                if worker_summary_col_widths and hasattr(self, 'worker_summary_tree'):
                    for col, width in worker_summary_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.worker_summary_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass
                # [NEW] Restore report column widths
                report_col_widths = config.get('report_col_widths', {})
                if report_col_widths and hasattr(self, 'report_tree'):
                    for col, width in report_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.report_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass
                
                if hasattr(self, '_loading_memos'):
                    del self._loading_memos
                    
                # Restore layout lock state
                if hasattr(self, 'btn_lock_layout'):
                    if self.layout_locked:
                        self.btn_lock_layout.config(text="🔒 배치 고정됨")
                        self.style.configure("Lock.TButton", foreground="black")
                    else:
                        self.btn_lock_layout.config(text="🔓 배치 수정 중")
                        self.style.configure("Lock.TButton", foreground="red")

                if config.get('daily_usage_sash_locked', False):
                    self.daily_usage_sash_locked = True
                    
                    if hasattr(self, 'btn_sash_lock'):
                        self.btn_sash_lock.config(text="🔒 경계 잠금됨")
                    
                    print("LOADED: Restored sash lock state")

                # Tab selection handled already at line 4531
                
                # Recreate Memos and Clones first (these must exist before they can be placed)
                self._loading_memos = []
                # Map class names to actual classes for recreation
                from site_apps.central.src.views.components import VehicleInspectionWidget
                class_map = {'Entry': ttk.Entry, 'Combobox': ttk.Combobox, 'VehicleInspectionWidget': VehicleInspectionWidget}
                
                draggable_geos = config.get('draggable_geometries', {})
                for key, geo in draggable_geos.items():
                    if key.startswith('clone_'):
                        self._loading_memos.append(key)
                        cls_name = geo.get('widget_class_name', 'Entry')
                        cls = class_map.get(cls_name, ttk.Entry)
                        label = geo.get('custom_label', "복제항목")
                        kwargs = geo.get('widget_kwargs', {})
                        m_list_key = geo.get('manage_list_key')
                        cont, w = self.create_draggable_container(self.entry_inner_frame, label, cls, key, manage_list_key=m_list_key, **kwargs)
                    # [UI REVISION] Legacy floating boxes (memo, checklist, vehicle) are no longer recreated 
                    # because they are now permanently embedded in the bottom dashboard.

                # 2. DELAYED RESTORATION: Restore complex states after UI mapped
                def delayed_restore():
                    self.root.update_idletasks()
                    
                    # 1. Restore window state and size
                    try:
                        if config.get('window_state'):
                            if config['window_state'] == 'zoomed':
                                self.root.state('zoomed')
                            else:
                                try:
                                    self.root.state(config['window_state'])
                                except: pass
                        
                        w = config.get('window_width')
                        h = config.get('window_height')
                        if w and h:
                            # Only set geometry if not zoomed
                            if self.root.state() != 'zoomed':
                                self.root.geometry(f"{w}x{h}")
                    except: pass

                    # 2. Restore resolution lock if active
                    if self.resolution_locked:
                        self.root.resizable(False, False)
                        if hasattr(self, 'btn_resolution_lock'):
                            self.btn_resolution_lock.config(text="🔒 해상도 고정됨")

                    # 3. Restore draggable components regardless of lock state
                    # [FIX] Always restore positions if they exist, so movement is persisted 
                    # even if the app was closed while 'unlocked'.
                    draggable_geos = config.get('draggable_geometries', {})
                    for key, geo in draggable_geos.items():
                        if key in self.draggable_items:
                            widget = self.draggable_items[key]
                            try:
                                # [LAYOUT FIX] 핵심 박스는 저장 좌표를 무시하고 기본 grid 배치 사용
                                if key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
                                    continue

                                if geo.get('hidden'):
                                    # [STABILITY] If it's a core widget, ignore hidden status and reset it
                                    if key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
                                        print(f"LOADER: Recovering hidden core widget: {key}")
                                        self.reset_widget_position(None, widget=widget)
                                        continue
                                        
                                    if widget.winfo_manager() == 'grid':
                                        widget.grid_forget()
                                    elif widget.winfo_manager() == 'place':
                                        widget.place_forget()
                                    continue
                                
                                if geo.get('custom_label') and hasattr(widget, '_label_widget'):
                                    widget._label_widget.config(text=geo['custom_label'])
                                
                                # [FIX] Explicitly un-grid before using place manager
                                if widget.winfo_manager() == 'grid':
                                    if not hasattr(widget, '_original_grid_info'):
                                        widget._original_grid_info = widget.grid_info()
                                    widget.grid_forget()
                                
                                self._ensure_placeholder(widget, width=geo.get('width'), height=geo.get('height'))
                                widget.lift()
                                
                                # [FIX] Enforce minimum dimensions to prevent invisible widgets
                                load_x = geo.get('x', 0)
                                load_y = geo.get('y', 0)
                                load_w = max(100, geo.get('width', 200))
                                load_h = max(50, geo.get('height', 100))

                                # [RECOVERY] 작업자 박스는 화면 밖/과소 복원을 방지
                                if key == 'workers_box_geometry':
                                    try:
                                        p_w = widget.master.winfo_width()
                                        p_h = widget.master.winfo_height()
                                        load_w = max(520, load_w)
                                        load_h = max(360, load_h)
                                        if p_w > 0:
                                            load_x = max(0, min(int(load_x), max(0, p_w - load_w)))
                                        else:
                                            load_x = max(0, int(load_x))
                                        if p_h > 0:
                                            load_y = max(0, min(int(load_y), max(0, p_h - load_h)))
                                        else:
                                            load_y = max(0, int(load_y))
                                    except Exception:
                                        load_x = max(0, int(load_x))
                                        load_y = max(0, int(load_y))
                                
                                widget.place(x=load_x, y=load_y, width=load_w, height=load_h)

                                # [VISUAL FIX] 핵심 박스는 기존 placeholder 잔상 제거 + 최소 필요 높이 보정
                                if key in ('ndt_usage_box_geometry', 'rtk_usage_box_geometry', 'workers_box_geometry'):
                                    try:
                                        self._remove_placeholder(widget)
                                    except Exception:
                                        pass
                                    try:
                                        widget.update_idletasks()
                                        if key == 'workers_box_geometry':
                                            min_w = max(520, widget.winfo_reqwidth() + 8)
                                            min_h = max(360, widget.winfo_reqheight() + 12)
                                        else:
                                            min_w = max(160, widget.winfo_reqwidth() + 4)
                                            min_h = max(90, widget.winfo_reqheight() + 6)
                                        cur_x = int(float(widget.place_info().get('x', load_x)))
                                        cur_y = int(float(widget.place_info().get('y', load_y)))
                                        cur_w = int(float(widget.place_info().get('width', load_w)))
                                        cur_h = int(float(widget.place_info().get('height', load_h)))
                                        if cur_w < min_w or cur_h < min_h:
                                            widget.place(x=cur_x, y=cur_y, width=max(cur_w, min_w), height=max(cur_h, min_h))
                                        if key == 'workers_box_geometry':
                                            widget.lift()
                                    except Exception:
                                        pass
                            except Exception as e:
                                print(f"Error placing widget {key}: {e}")
                                
                    # 3.5 [LAYOUT FIX] 핵심 박스는 항상 기본 grid 위치로 복구
                    for key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
                        if key not in self.draggable_items:
                            continue
                        widget = self.draggable_items[key]
                        try:
                            self.reset_widget_position(None, widget=widget)
                            self._remove_placeholder(widget)
                            widget.lift()
                        except Exception as e:
                            print(f"LOADER: Failed to reset core widget {key}: {e}")

                    # 3.8 [SAFEGUARD] 저장 버튼 강제 복구 (화면 밖/숨김/겹침 방지)
                    if 'save_btn_geometry' in self.draggable_items:
                        sbox = self.draggable_items['save_btn_geometry']
                        try:
                            sbox.update_idletasks()
                            smgr = sbox.winfo_manager()
                            sw = sbox.winfo_width()
                            sh = sbox.winfo_height()

                            # Hidden/unmanaged or too small => reset to grid default first
                            if smgr == '' or sw < 80 or sh < 20:
                                self.reset_widget_position(None, widget=sbox)
                                smgr = sbox.winfo_manager()

                            # If placed, clamp into visible area and keep above siblings
                            if smgr == 'place':
                                pi = sbox.place_info()
                                sx = int(float(pi.get('x', 10)))
                                sy = int(float(pi.get('y', 340)))
                                sw = int(float(pi.get('width', max(120, sbox.winfo_reqwidth() or 120))))
                                sh = int(float(pi.get('height', max(28, sbox.winfo_reqheight() or 28))))
                                p_w = sbox.master.winfo_width()
                                p_h = sbox.master.winfo_height()
                                sw = max(120, sw)
                                sh = max(28, sh)
                                if p_w > 0:
                                    sx = max(0, min(sx, max(0, p_w - sw)))
                                if p_h > 0:
                                    sy = max(0, min(sy, max(0, p_h - sh)))
                                sbox.place(x=sx, y=sy, width=sw, height=sh)

                            sbox.lift()
                        except Exception:
                            pass

                    # 4. Restore sash positions
                    def apply_sashes():
                        try:
                            # Daily usage sash
                            daily_sash = config.get('daily_usage_sash_pos')
                            if daily_sash is not None and hasattr(self, 'daily_usage_paned'):
                                total_h = self.daily_usage_paned.winfo_height()
                                if total_h > 0:
                                    min_pos = max(int(total_h * 0.1), 460)
                                    max_pos = int(total_h * 0.9)
                                    safe_pos = int(daily_sash)
                                    if safe_pos < min_pos:
                                        safe_pos = min_pos
                                    if safe_pos > max_pos:
                                        safe_pos = max_pos
                                    getattr(self.daily_usage_paned, "sashpos", lambda *args: 500)(0, safe_pos)
                                else:
                                    getattr(self.daily_usage_paned, "sashpos", lambda *args: 500)(0, int(daily_sash))
                            
                            # History sash
                            history_sash = config.get('daily_history_sash_pos')
                            if history_sash is not None and hasattr(self, 'daily_history_paned'):
                                self.daily_history_paned.sashpos(0, int(history_sash))
                                
                            # Bottom dashboard sashes
                            bottom_sashes = config.get('bottom_dashboard_sashes', [])
                            if bottom_sashes and len(bottom_sashes) >= 2 and hasattr(self, 'bottom_dashboard') and isinstance(self.bottom_dashboard, ttk.PanedWindow):
                                try:
                                    # Ensure widget is fully drawn before setting sashes
                                    self.bottom_dashboard.update_idletasks()
                                    total_w = self.bottom_dashboard.winfo_width()
                                    
                                    if total_w > 300: # Make sure window is mapped and has a reasonable width
                                        # Clamp sash 1: at least 100px from left, at most 200px from right edge
                                        s1 = max(100, min(int(bottom_sashes[0]), total_w - 200))
                                        # Clamp sash 2: at least 100px from sash 1, at most 50px from right edge
                                        s2 = max(s1 + 100, min(int(bottom_sashes[1]), total_w - 50))
                                        
                                        self.bottom_dashboard.sashpos(0, s1)
                                        self.bottom_dashboard.sashpos(1, s2)
                                    else:
                                        self.bottom_dashboard.sashpos(0, int(bottom_sashes[0]))
                                        self.bottom_dashboard.sashpos(1, int(bottom_sashes[1]))
                                except: pass
                                
                            # If sash lock is active (check config heavily)
                            config_locked = config.get('daily_usage_sash_locked', False)
                            if config_locked:
                                self.daily_usage_sash_locked = True
                                if hasattr(self, 'btn_sash_lock'):
                                    self.btn_sash_lock.config(text="🔒 경계 고정됨")
                                    self.btn_sash_lock.configure(style="SashLock.TButton")
                                self._start_sash_monitor()
                            elif self.daily_usage_sash_locked: # If self says locked but config didn't (fallback)
                                self._start_sash_monitor()
                        except: pass

                    self.root.after(100, apply_sashes)
                    self.root.after(800, apply_sashes)

                    # 5. Refresh canvas scroll region
                    if hasattr(self, 'entry_inner_frame'):
                        self._ensure_canvas_scroll_region()
                    
                    # 6. Final UI refresh
                    for l_key in ['users', 'sites', 'equipments', 'warehouses', 'worktimes']:
                        self.refresh_ui_for_list_change(l_key)

                    # 7. Mark as ready for future saves AFTER all restoration is done
                    def finalize_loading():
                        try:
                            print("[STARTUP] Finalizing loading...")
                            self.is_ready = True
                            
                            # [FIX] Forcibly show and deiconify window to prevent "closing downwards" or disappearing
                            print("[STARTUP] Ensuring window visibility...")
                            try:
                                self.root.deiconify()
                                self.root.lift()
                                self.root.focus_force()
                            except: pass
                            
                            # [V13_RESET_ON_STARTUP] Ensure Daily Usage form starts BLANK as requested
                            print("[STARTUP] Resetting forms...")
                            try:
                                self.clear_daily_usage_form_all()
                            except: pass
                            
                            # [NEW] Trigger column auto-hide on both tabs immediately after startup
                            print("[STARTUP] Refreshing views...")
                            try:
                                self.root.after(100, lambda: self.update_daily_usage_view())
                                self.root.after(200, lambda: self.update_monthly_usage_view())
                                self.root.after(300, lambda: self.update_budget_site_view()) 
                                # Removed hardcoded default tab selection so the user's saved tab is preserved
                            except Exception as e:
                                print(f"[STARTUP] View refresh error: {e}")
                                
                            print("APP READY: State restoration complete.")
                        except Exception as startup_err:
                            print(f"[CRITICAL STARTUP ERROR] {startup_err}")
                            messagebox.showerror("Startup Error", f"An error occurred during startup: {startup_err}\n\n{traceback.format_exc()}")
                    
                    print("[STARTUP] Scheduling finalization...")
                    self.root.after(500, finalize_loading) 
                
                # Execute delayed restoration
                try:
                    self.root.after(300, delayed_restore)
                except Exception as outer_err:
                    messagebox.showerror("Initialization Error", f"Critical error during initialization: {outer_err}")

        except Exception as e:
            print(f"Failed to load tab config: {e}")
            # Ensure we eventually become ready even if load failed, to allow new saves
            self.root.after(2000, lambda: setattr(self, 'is_ready', True))
    
    def on_closing(self, *args, **kwargs):
        from site_apps.central.src.controllers.event_controller import on_closing_impl
        return on_closing_impl(self, *args, **kwargs)
    
    def export_stock_to_excel(self):
        """Export current stock data to Excel"""
        try:
            # Get current filtered data from treeview
            stock_data = []
            
            for item in self.stock_tree.get_children():
                values = self.stock_tree.item(item, 'values')
                stock_data.append({
                    'No.': values[0],
                    '회사코드': values[1],
                    '관리품번': values[2],
                    '품목명': values[3],
                    'SN': values[4],
                    '창고': values[5],
                    '모델명': values[6],
                    '규격': values[7],
                    '품목군코드': values[8],
                    '공급업체': values[9],
                    '제조사': values[10],
                    '제조국': values[11],
                    '가격': values[12],
                    '관리단위': values[13],
                    '수량': values[14],
                    '재고하한': values[15]
                })
            
            if not stock_data:
                messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
                return
            
            # Prepare filename with current date
            current_date = datetime.datetime.now().strftime('%Y%m%d')
            filename = f"재고현황_{current_date}.xlsx"
            
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=filename,
                title="재고 현황 저장",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if save_path:
                stock_df = pd.DataFrame(stock_data)
                stock_df = self.clean_df_export(stock_df)
                self.save_df_to_excel_autofit(stock_df, save_path, "재고현황")
                messagebox.showinfo("완료", f"재고 현황이 저장되었습니다.\n저장 위치: {save_path}")
                
        except Exception as e:
            messagebox.showerror("오류", f"내보내기 실패: {e}")

    def enable_tree_column_drag(self, tree, context_menu_handler=None):
        """Ultra-robust drag-and-drop column reordering using root-level event tracking"""
        tree._drag_info = {"col_id": None, "col_name": None, "start_x": 0, "active": False, "ghost": None}
        
        def on_press(event):
            region = tree.identify_region(event.x, event.y)
            if region == "heading":
                col_id = tree.identify_column(event.x)
                col_name = self._get_column_name_from_id(tree, col_id)
                if col_name:
                    tree._drag_info.update({
                        "col_id": col_id,
                        "col_name": col_name,
                        "start_x": event.x_root,
                        "active": False,
                        "motion_id": self.root.bind("<B1-Motion>", on_motion, add="+"),
                        "release_id": self.root.bind("<ButtonRelease-1>", on_release, add="+")
                    })
                
        def on_motion(event):
            di = tree._drag_info
            if not di["col_name"]: return
            
            # Start drag if moved > 5 pixels
            if not di["active"] and abs(event.x_root - di["start_x"]) > 5:
                di["active"] = True
                tree.config(cursor="fleur")
                
                # Visual ghost label
                try:
                    if di["ghost"]: di["ghost"].destroy()
                    di["ghost"] = tk.Toplevel(self.root)
                    di["ghost"].overrideredirect(True)
                    di["ghost"].attributes("-alpha", 0.8)
                    di["ghost"].attributes("-topmost", True)
                    
                    lbl = tk.Label(di["ghost"], text=f" ↔ {di['col_name']} 이동 중 ", 
                                   bg="#424242", fg="white", relief="raised", borderwidth=1, 
                                   padx=8, pady=4, font=('Malgun Gothic', 10, 'bold'))
                    lbl.pack()
                except: pass
            
            if di["active"] and di["ghost"]:
                try:
                    # Follow mouse with offset to stay visible
                    di["ghost"].geometry(f"+{event.x_root + 20}+{event.y_root + 10}")
                except: pass

        def on_release(event):
            di = tree._drag_info
            if di["active"] and di["col_name"]:
                # Convert screen X back to tree-relative X
                tree_x = event.x_root - tree.winfo_rootx()
                target_id = tree.identify_column(tree_x)
                
                if target_id and target_id != di["col_id"]:
                    self._reorder_tree_columns(tree, di["col_id"], target_id)
            
            # Cleanup all bindings and state
            tree.config(cursor="")
            if di["ghost"]:
                try: di["ghost"].destroy()
                except: pass
                di["ghost"] = None
            
            # Unbind specific root-level drag listeners securely
            try:
                if di.get("motion_id"):
                    self.root.unbind("<B1-Motion>", di["motion_id"])
                if di.get("release_id"):
                    self.root.unbind("<ButtonRelease-1>", di["release_id"])
            except: pass
            
            di.update({"col_id": None, "col_name": None, "active": False, "motion_id": None, "release_id": None})

        # Primary entry point: tree-level press
        tree.bind("<Button-1>", on_press, add="+")
        # Direct right-click menu as reliable alternative
        if context_menu_handler is not None:
            tree.bind("<Button-3>", context_menu_handler, add="+")
        else:
            tree.bind("<Button-3>", lambda e: self._show_heading_context_menu(e, tree), add="+")

    def _show_heading_context_menu(self, event, tree):
        """Show context menu on Treeview header to move columns"""
        region = tree.identify_region(event.x, event.y)
        if region == "heading":
            column_id = tree.identify_column(event.x)
            
            # Identify the column name from visual ID
            col_name = self._get_column_name_from_id(tree, column_id)
            if not col_name: return

            menu = tk.Menu(self.root, tearoff=0)
            menu.add_command(label=f"⬅️ '{col_name}' 왼쪽으로 이동", 
                             command=lambda: self._move_column_visual(tree, column_id, -1))
            menu.add_command(label=f"➡️ '{col_name}' 오른쪽으로 이동", 
                             command=lambda: self._move_column_visual(tree, column_id, 1))
            menu.add_separator()
            menu.add_command(label="⚙️ 컬럼 관리(숨기기/보이기)...", command=self.show_column_visibility_dialog)
            
            menu.post(event.x_root, event.y_root)

    def _show_generic_tree_heading_context_menu(self, event, tree):
        """Generic header context menu for trees that support only column move."""
        try:
            if tree.identify_region(event.x, event.y) != "heading":
                return
            column_id = tree.identify_column(event.x)
            col_name = self._get_column_name_from_id(tree, column_id)
            if not col_name:
                return
        except Exception:
            return

        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label=f"⬅️ '{col_name}' 왼쪽으로 이동",
                         command=lambda: self._move_column_visual(tree, column_id, -1))
        menu.add_command(label=f"➡️ '{col_name}' 오른쪽으로 이동",
                         command=lambda: self._move_column_visual(tree, column_id, 1))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _get_column_name_from_id(self, tree, column_id):
        """Helper to get actual column name from visual ID like #1"""
        try:
            if not column_id or not column_id.startswith("#"): return None
            # Extract N from #N
            vis_idx = int(column_id[1:]) - 1
            
            # Get displaycolumns accurately
            disp_tuple = tree.cget("displaycolumns")
            if not disp_tuple or disp_tuple == ("#all",) or disp_tuple == ["#all"] or disp_tuple == "":
                visible_names = list(tree.cget("columns"))
            else:
                visible_names = list(disp_tuple)
            
            if 0 <= vis_idx < len(visible_names):
                return visible_names[vis_idx]
        except: pass
        return None

    def _move_column_visual(self, tree, source_id, direction):
        """Move a column left (-1) or right (1) based on its current position"""
        col_name = self._get_column_name_from_id(tree, source_id)
        if not col_name: return
        
        try:
            disp_tuple = tree.cget("displaycolumns")
            if not disp_tuple or disp_tuple == ("#all",) or disp_tuple == ["#all"] or disp_tuple == "":
                visible_names = list(tree.cget("columns"))
            else:
                visible_names = list(disp_tuple)

            if col_name not in visible_names: return
            
            src_idx = visible_names.index(col_name)
            tgt_idx = src_idx + direction
            
            if 0 <= tgt_idx < len(visible_names):
                visible_names.pop(src_idx)
                visible_names.insert(tgt_idx, col_name)
                tree["displaycolumns"] = visible_names
                
                # Update persistence
                if tree == self.daily_usage_tree:
                    self.manual_visible_cols = visible_names
                    self.save_tab_config()
                elif tree == self.monthly_usage_tree:
                    self.monthly_visible_cols = visible_names
                    self.save_tab_config()
                elif hasattr(self, 'budget_view_tree') and tree == self.budget_view_tree:
                    self.budget_view_visible_cols = visible_names
                    self.save_tab_config()
        except: pass

    def _reorder_tree_columns(self, tree, source_id, target_id):
        """Internal helper to swap displaycolumns for drag-and-drop"""
        src_name = self._get_column_name_from_id(tree, source_id)
        tgt_name = self._get_column_name_from_id(tree, target_id)
        
        if not src_name or not tgt_name or src_name == tgt_name: return
        
        try:
            disp_tuple = tree.cget("displaycolumns")
            if not disp_tuple or disp_tuple == ("#all",) or disp_tuple == ["#all"] or disp_tuple == "":
                visible_names = list(tree.cget("columns"))
            else:
                visible_names = list(disp_tuple)
            
            if src_name in visible_names and tgt_name in visible_names:
                src_idx = visible_names.index(src_name)
                tgt_idx = visible_names.index(tgt_name)
                
                # Strategic insertion: move before or after target 
                # (Simple swap approach: just pop and insert)
                visible_names.pop(src_idx)
                visible_names.insert(tgt_idx, src_name)
                
                tree["displaycolumns"] = visible_names
                
                # Update persistence
                if tree == self.daily_usage_tree:
                    self.manual_visible_cols = visible_names
                    self.save_tab_config()
                elif tree == self.monthly_usage_tree:
                    self.monthly_visible_cols = visible_names
                    self.save_tab_config()
                elif hasattr(self, 'budget_view_tree') and tree == self.budget_view_tree:
                    self.budget_view_visible_cols = visible_names
                    self.save_tab_config()
        except: pass


    def treeview_sort_column(self, tv, col, reverse):
        """Sort treeview contents when a column header is clicked"""
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        
        # Separate the 'Total' row (tagged 'total') from sorting
        data_rows = []
        total_row = None
        
        for val, k in l:
            # Check tags for this item
            tags = tv.item(k, 'tags')
            # If tags is a tuple/list, check if 'total' is in it. 
            # If it's a string (though usually tuple), check equality or contains.
            if tags and ('total' in tags or tags == 'total'):
                total_row = (val, k)
            else:
                data_rows.append((val, k))
                
        # Helper for numeric conversion
        def convert(val):
            try:
                # Remove common non-numeric chars
                s = str(val).replace(',', '').replace('시간', '').replace('Hrs', '').replace('원', '').replace(' ', '').replace('(', '').replace(')', '')
                if not s: return 0.0
                return float(s)
            except ValueError:
                return str(val).lower() # Default to string sort

        try:
            data_rows.sort(key=lambda t: convert(t[0]), reverse=reverse)
        except Exception:
            data_rows.sort(key=lambda t: t[0].lower(), reverse=reverse)

        # Rearrange items in sorted positions
        for index, (val, k) in enumerate(data_rows):
            tv.move(k, '', index)
            
        # Ensure Total row is always at the bottom
        if total_row:
             tv.move(total_row[1], '', 'end')

        # Reverse sort next time
        tv.heading(col, command=lambda: self.treeview_sort_column(tv, col, not reverse))

    def switch_to_daily_usage_details(self):
        """Switch to Daily Usage tab and apply current filters from Sales tab"""
        # Sync filters before switching

        site = self.cb_sales_filter_site.get().strip()
        start = self.sales_start_date.get_date()
        end = self.sales_end_date.get_date()
        
        # 1. Switch Tab (index 5)
        self.notebook.select(5)
        
        # [NEW] Force focus transition
        if hasattr(self, 'daily_usage_tree'):
            self.daily_usage_tree.focus_set()

        # 2. Sync Filters
        if hasattr(self, 'cb_daily_filter_site'):
            self.cb_daily_filter_site.set(site)
        
        if hasattr(self, 'ent_daily_start_date'):
            self.ent_daily_start_date.set_date(start)
            
        if hasattr(self, 'ent_daily_end_date'):
            self.ent_daily_end_date.set_date(end)
            
        # 3. Trigger Search
        self.update_daily_usage_view()

    def save_all_daily_usage_changes(self):
        """Unified save button for the inquiry list to satisfy spreadsheet-style workflow"""
        try:
            if self.save_data():
                # Provide clear feedback that ALL changes in the app are saved
                messagebox.showinfo("저장 완료", "모든 변경 사항이 마스터 파일(Material_Inventory.xlsx)에 안전하게 저장되었습니다.")
                # Refresh view to ensure everything is synchronized
                self.update_daily_usage_view()
                self.refresh_inquiry_filters()
        except Exception as e:
            messagebox.showerror("저장 오류", f"데이터 저장 중 오류가 발생했습니다: {e}")

    def open_detached_monthly_usage_view(self):
        """월별 집계 조회를 별도의 팝업창(모니터링 창)으로 엽니다. (메인 화면과 100% 동일한 구성)"""
        # 기존 창이 있으면 포커스만 이동
        if 'monthly' in self.detached_windows and self.detached_windows['monthly']['window'].winfo_exists():
            self.detached_windows['monthly']['window'].lift()
            return

        popup = tk.Toplevel(self.root)
        popup.title("📊 월별 집계 모니터링 (팝업)")
        popup.geometry("1600x900")
        
        main_frame = ttk.Frame(popup, padding=10)
        main_frame.pack(expand=True, fill='both')
        
        # 상단 정보/필터 안내
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill='x', pady=(0, 5))
        
        # [NEW] Popup Search Filters
        filter_row = ttk.Frame(info_frame)
        filter_row.pack(side='left', fill='x', expand=True)

        def sync_and_refresh(event=None):
            # Mirror popup filters to main window
            self.cb_filter_year.set(p_cb_year.get())
            self.cb_filter_month.set(p_cb_month.get())
            self.cb_filter_site_monthly.set(p_cb_site.get())
            self.cb_filter_material_monthly.set(p_cb_mat.get())
            self.update_monthly_usage_view()

        ttk.Label(filter_row, text="연도:").pack(side='left', padx=2)
        p_cb_year = ttk.Combobox(filter_row, values=self.cb_filter_year['values'], width=8)
        p_cb_year.pack(side='left', padx=2)
        p_cb_year.set(self.cb_filter_year.get())
        p_cb_year.bind("<<ComboboxSelected>>", sync_and_refresh)

        ttk.Label(filter_row, text="월:").pack(side='left', padx=2)
        p_cb_month = ttk.Combobox(filter_row, values=self.cb_filter_month['values'], width=6)
        p_cb_month.pack(side='left', padx=2)
        p_cb_month.set(self.cb_filter_month.get())
        p_cb_month.bind("<<ComboboxSelected>>", sync_and_refresh)

        ttk.Label(filter_row, text="현장:").pack(side='left', padx=2)
        p_cb_site = ttk.Combobox(filter_row, values=self.cb_filter_site_monthly['values'], width=12)
        p_cb_site.pack(side='left', padx=2)
        p_cb_site.set(self.cb_filter_site_monthly.get())
        p_cb_site.bind("<<ComboboxSelected>>", sync_and_refresh)

        ttk.Label(filter_row, text="품목:").pack(side='left', padx=2)
        p_cb_mat = ttk.Combobox(filter_row, values=self.cb_filter_material_monthly['values'], width=18)
        p_cb_mat.pack(side='left', padx=2)
        p_cb_mat.set(self.cb_filter_material_monthly.get())
        p_cb_mat.bind("<<ComboboxSelected>>", sync_and_refresh)

        ttk.Button(filter_row, text="조회", width=6, command=sync_and_refresh).pack(side='left', padx=5)
        
        ttk.Button(filter_row, text="🔄 새로고침", command=sync_and_refresh).pack(side='right', padx=5)
        
        paned = ttk.PanedWindow(main_frame, orient="vertical")
        paned.pack(expand=True, fill='both')
        
        # 1. 메인 트리뷰 (상단)
        tree_frame = ttk.Frame(paned)
        paned.add(tree_frame, weight=3) 
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        columns = ('연도', '월', '현장', '구분', '작업자', '작업시간', 'OT시간', 'OT금액', 'OT1', 'OT2', 'OT3', 'OT4', 'OT5', 'OT6', 'OT7', 'OT8', 'OT9', 'OT10', 
                   '수량', '단가', '출장비', '일식', '검사비', '품목명', '센터미스', '농도', '마킹미스', '필름마크', 
                   '취급부주의', '고객불만', '기타', 'RTK총계', '형광자분', '흑색자분', '백색페인트', '침투제', '세척제', '현상제', '형광침투제', '비고', '(Full작업자)')
        
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree['displaycolumns'] = [c for c in columns if c != '(Full작업자)']
        
        vsb.config(command=tree.yview); hsb.config(command=tree.xview)
        
        col_widths = {'연도': 90, '월': 70, '현장': 140, '구분': 100, '작업자': 100, '작업시간': 100, 'OT시간': 100, 'OT금액': 110, '품목명': 200, '비고': 220}
        for col in columns:
            tree.heading(col, text=col, command=lambda c=col: self.treeview_sort_column(tree, c, False))
            tree.column(col, width=col_widths.get(col, 100), anchor='center', stretch=False)
        
        tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1); tree_frame.grid_columnconfigure(0, weight=1)

        # [NEW] 인터랙티브 기능 복제
        self.enable_tree_column_drag(tree)
        tree.tag_configure('total', background='#E8F4F8', font=('Arial', 12, 'bold'))
        tree.bind("<Button-1>", lambda e: self.show_worker_popup(e, tree), add="+")
        tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())

        # 2. 현장별 요약 (중단 - 메인 UI는 3단 수직 분할)
        site_frame = ttk.LabelFrame(paned, text="현장별 누계")
        paned.add(site_frame, weight=1)
        site_cols = ('현장', '검사방법', '품목명', '수량', '검사비', '출장비', '제경비', '기술료', '환산물량', '재료비', '인건비', '형광자분', '흑색자분', '백색페인트', 
                     '침투제', '세척제', '현상제', '형광침투제', '센터미스', '농도', '마킹미스', '필름마크', '취급부주의', '고객불만', '기타', 'RTK총계')
        site_tree = ttk.Treeview(site_frame, columns=site_cols, show='headings')
        site_vsb = ttk.Scrollbar(site_frame, orient="vertical", command=site_tree.yview)
        site_tree.configure(yscrollcommand=site_vsb.set)
        for col in site_cols:
            site_tree.heading(col, text=col, command=lambda c=col: self.treeview_sort_column(site_tree, c, False))
            w = 120 if col in ['현장', '검사방법', '품목명'] else (100 if col in ['검사비', '출장비', '제경비', '기술료', '재료비', '인건비'] else 80)
            site_tree.column(col, width=w, anchor='center', stretch=False)
        site_tree.pack(side='left', expand=True, fill='both')
        site_vsb.pack(side='right', fill='y')
        
        self.enable_tree_column_drag(site_tree, context_menu_handler=lambda e: self._show_generic_tree_heading_context_menu(e, site_tree))
        site_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())
        
        # 3. 작업자별 요약 (하단)
        worker_frame = ttk.LabelFrame(paned, text="작업자별 누계")
        paned.add(worker_frame, weight=1)
        worker_cols = ('작업자', '총공수', '연장(시간)', '야간(시간)', '휴일(시간)', '총OT(시간)', '연장(금액)', '야간(금액)', '휴일(금액)', '총OT(금액)')
        worker_tree = ttk.Treeview(worker_frame, columns=worker_cols, show='headings')
        worker_vsb = ttk.Scrollbar(worker_frame, orient="vertical", command=worker_tree.yview)
        worker_tree.configure(yscrollcommand=worker_vsb.set)
        worker_widths = {'작업자': 100, '총공수': 70, '총OT(시간)': 80, '총OT(금액)': 100}
        for col in worker_cols:
            worker_tree.heading(col, text=col, command=lambda c=col: self.treeview_sort_column(worker_tree, c, False))
            worker_tree.column(col, width=worker_widths.get(col, 80), anchor='center', stretch=False)
        worker_tree.pack(side='left', expand=True, fill='both')
        worker_vsb.pack(side='right', fill='y')

        self.enable_tree_column_drag(worker_tree, context_menu_handler=lambda e: self._show_generic_tree_heading_context_menu(e, worker_tree))
        worker_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())

        # [NEW] 메인 UI의 분할 배율 동기화
        def sync_sash():
            try:
                main_h = self.monthly_paned.winfo_height()
                if main_h > 0:
                    popup_h = paned.winfo_height()
                    if popup_h > 0:
                        s1 = self.monthly_paned.sashpos(0)
                        s2 = self.monthly_paned.sashpos(1)
                        paned.sashpos(0, int(s1 * popup_h / main_h))
                        paned.sashpos(1, int(s2 * popup_h / main_h))
            except: pass
        popup.after(500, sync_sash) # UI 렌더링 후 적용

        # 상태 등록
        self.detached_windows['monthly'] = {
            'window': popup,
            'tree': tree,
            'site_tree': site_tree,
            'worker_tree': worker_tree,
            'filters': {
                'year': p_cb_year,
                'month': p_cb_month,
                'site': p_cb_site,
                'mat': p_cb_mat
            }
        }
        
        def on_close():
            if 'monthly' in self.detached_windows:
                del self.detached_windows['monthly']
            popup.destroy()
        popup.protocol("WM_DELETE_WINDOW", on_close)
        
        # 데이터 복사
        self.update_monthly_usage_view()
        
        # [FIXED] Extraction using standardized indices (Year=0, Month=1, Site=2)
        def on_popup_select(event):
            selection = tree.selection()
            if not selection: return
            item = selection[0]
            try:
                vals = tree.item(item, 'values')
                if not vals: return
                tags = tree.item(item, 'tags')
                
                if not hasattr(self, 'current_monthly_df'): return
                
                if 'total' in tags:
                    subset = self.current_monthly_df
                else:
                    # Robust extraction
                    try:
                        y = int(vals[0]) 
                        m = int(vals[1])
                        s = str(vals[2]).strip()
                        
                        # Material name is at index 22 in our formatted tuple
                        mat_name = str(vals[22]).strip() if len(vals) > 22 else None
                        
                        mask = (self.current_monthly_df['Year'] == y) & \
                               (self.current_monthly_df['Month'] == m) & \
                               (self.current_monthly_df['Site'] == s)
                        
                        if mat_name:
                            ids = self.materials_df[self.materials_df['품목명'] == mat_name]['MaterialID'].tolist()
                            if ids:
                                mask = mask & (self.current_monthly_df['MaterialID'].isin(ids))
                        
                        subset = self.current_monthly_df[mask]
                    except:
                        subset = pd.DataFrame()
                
                # Update summaries
                self._populate_monthly_summary_trees(subset)
            except Exception as e:
                print(f"DEBUG: Detached selection error: {e}")

        tree.bind("<<TreeviewSelect>>", on_popup_select)
        
        if hasattr(self, 'current_monthly_df'):
            self._populate_monthly_summary_trees(self.current_monthly_df)

    def open_detached_budget_view(self):
        """공사실행예산서 탭을 별도의 팝업창으로 엽니다. (메인 화면과 유사한 KPI/목록 구성)"""
        if 'budget' in self.detached_windows and self.detached_windows['budget']['window'].winfo_exists():
            self.detached_windows['budget']['window'].lift()
            return

        popup = tk.Toplevel(self.root)
        popup.title("📋 공사 실행예산서 모니터링 (팝업)")
        popup.geometry("1400x800")
        
        main_frame = ttk.Frame(popup, padding=10)
        main_frame.pack(expand=True, fill='both')
        
        # 1. KPI Panel (메인 UI 그대로 재현)
        kpi_frame = tk.Frame(main_frame, bg='#f8fafc', height=80)
        kpi_frame.pack(fill='x', pady=(0, 10))
        kpi_frame.pack_propagate(False)
        
        def create_kpi_card(parent, title, color):
            frame = tk.Frame(parent, bg='white', highlightbackground='#e2e8f0', highlightthickness=1)
            frame.pack(side='left', expand=True, fill='both', padx=5, pady=5)
            tk.Label(frame, text=title, font=('Malgun Gothic', 9), fg='#64748b', bg='white').pack(pady=(5, 2))
            lbl = tk.Label(frame, text="0", font=('Malgun Gothic', 16, 'bold'), fg=color, bg='white')
            lbl.pack(pady=(0, 5))
            return lbl

        lbl_rev = create_kpi_card(kpi_frame, "총 매출액", "#2563eb")
        lbl_cost = create_kpi_card(kpi_frame, "총 실행원가", "#dc2626")
        lbl_profit = create_kpi_card(kpi_frame, "총 기대이익", "#059669")
        lbl_margin = create_kpi_card(kpi_frame, "평균 이익률", "#7c3aed")
        
        # 2. 예산 목록 트리뷰
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(expand=True, fill='both')
        
        vsb = ttk.Scrollbar(tree_frame, orient='vertical')
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal')
        
        columns = ('Site', 'Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit', 'TargetMargin', 'ActualMargin', 'Note')
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        col_map = {
            'Site': ('현장명', 150), 'Revenue': ('매출액(목표)', 120), 'UnitPrice': ('단가', 100),
            'LaborCost': ('노무비', 120), 'MaterialCost': ('자재비', 120), 'Expense': ('경비', 120),
            'OutsourceCost': ('외주비', 120), 'Profit': ('기대이익(목표)', 120), 'TargetMargin': ('목표이익률', 100), 'ActualMargin': ('현재이익률', 100), 'Note': ('비고', 200)
        }
        for col in columns:
            head, width = col_map[col]
            tree.heading(col, text=head, command=lambda c=col: self.treeview_sort_column(tree, c, False))
            tree.column(col, width=width, anchor='center', stretch=False)
            
        tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1); tree_frame.grid_columnconfigure(0, weight=1)
        
        vsb.config(command=tree.yview); hsb.config(command=tree.xview)
        
        # [NEW] 인터랙티브 기능 추가
        self.enable_tree_column_drag(tree)

        self.detached_windows['budget'] = {
            'window': popup,
            'tree': tree,
            'lbl_rev': lbl_rev,
            'lbl_cost': lbl_cost,
            'lbl_profit': lbl_profit,
            'lbl_margin': lbl_margin
        }
        
        def on_close():
            if 'budget' in self.detached_windows:
                del self.detached_windows['budget']
            popup.destroy()
        popup.protocol("WM_DELETE_WINDOW", on_close)
        
        self._refresh_detached_budget()

    def _refresh_detached_budget(self):
        """팝업창의 예산 리스트와 KPI를 최신 데이터로 갱신"""
        if 'budget' not in self.detached_windows: return
        detached = self.detached_windows['budget']
        tree = detached['tree']
        
        for item in tree.get_children(): tree.delete(item)
        
        if self.budget_df.empty:
            for lbl in [detached['lbl_rev'], detached['lbl_cost'], detached['lbl_profit']]: lbl.config(text="0")
            detached['lbl_margin'].config(text="0.0%")
            return
        
        total_rev = total_cost = total_profit = 0
        
        for _, row in self.budget_df.iterrows():
            rev = pd.to_numeric(row.get('Revenue', 0), errors='coerce') or 0
            unit_p = pd.to_numeric(row.get('UnitPrice', 0), errors='coerce') or 0
            lab = pd.to_numeric(row.get('LaborCost', 0), errors='coerce') or 0
            mat = pd.to_numeric(row.get('MaterialCost', 0), errors='coerce') or 0
            exp = pd.to_numeric(row.get('Expense', 0), errors='coerce') or 0
            out = pd.to_numeric(row.get('OutsourceCost', 0), errors='coerce') or 0
            prof = pd.to_numeric(row.get('Profit', 0), errors='coerce') or 0
            
            a_rev = pd.to_numeric(row.get('Actual_Revenue', 0), errors='coerce') or 0
            a_prof = pd.to_numeric(row.get('Actual_Profit', 0), errors='coerce') or 0
            
            total_rev += rev
            total_cost += (lab + mat + exp + out)
            total_profit += prof
            
            target_margin = (prof / rev * 100) if rev > 0 else 0
            actual_margin = (a_prof / a_rev * 100) if a_rev > 0 else 0
            
            tree.insert('', 'end', values=(
                row.get('Site', ''),
                f"{rev:,.0f}",
                f"{unit_p:,.0f}",
                f"{lab:,.0f}",
                f"{mat:,.0f}",
                f"{exp:,.0f}",
                f"{out:,.0f}",
                f"{prof:,.0f}",
                f"{target_margin:.1f}%",
                f"{actual_margin:.1f}%",
                row.get('Note', '')
            ))
            
        # KPI Update
        detached['lbl_rev'].config(text=f"{total_rev:,.0f}")
        detached['lbl_cost'].config(text=f"{total_cost:,.0f}")
        detached['lbl_profit'].config(text=f"{total_profit:,.0f}")
        total_margin = (total_profit / total_rev * 100) if total_rev > 0 else 0
        detached['lbl_margin'].config(text=f"{total_margin:.1f}%")


    def cleanup_auto_transactions(self):
        """Remove existing 'Automatic Deduction' transactions for non-consumable equipment (PAUT, Yokes, etc.)"""
        if self.transactions_df.empty: 
            messagebox.showinfo("알림", "처리할 입출고 기록이 없습니다.")
            return
        
        confirm = messagebox.askyesno("클린업 확인", 
            "현재 저장된 '자동 차감' 기록 중 장비류(PAUT, Yoke 등)에 해당하는 내역만 골라 삭제하시겠습니까?\n\n"
            "※ 필름, 약품 등 소모품 기록은 유지됩니다.")
        if not confirm: return

        initial_count = len(self.transactions_df)
        
        def is_bad_auto(row):
            note = str(row.get('Note', ''))
            # [FIX] Match both English and potentially broken encoding versions of "(자동 차감)"
            # (ڵ ) is a common corruption of (자동 차감) when read as different encoding
            is_auto = "(자동 차감)" in note or "(ڵ )" in note or "(\ucac0 \u0020)" in note or "(\u00e1\u00b6\u00bd\u0020)" in note

            if not is_auto:
                return False
            
            mat_id = row.get('MaterialID')
            mat_name = self.get_material_display_name(mat_id)
            
            # If it's NOT a consumable, it's a candidate for removal
            # _is_consumable_material handles empty method by returning False for equipment keywords
            return not self._is_consumable_material(mat_name, "")

        mask = self.transactions_df.apply(is_bad_auto, axis=1)
        self.transactions_df = self.transactions_df[~mask]
        
        removed_count = initial_count - len(self.transactions_df)
        if removed_count > 0:
            if self.save_data():
                messagebox.showinfo("클린업 완료", f"장비류(PAUT, Yoke 등)의 자동 차감 내역 {removed_count}건을 삭제했습니다.")
                self.update_stock_view()
                self.update_transaction_view()
        else:
            messagebox.showinfo("알림", "정리할 대상이 없습니다.")

if __name__ == "__main__":
    root = tk.Tk()
    app = MaterialManager(root)
    try:
        root.mainloop()
    except KeyboardInterrupt:
        try:
            root.destroy()
        except Exception:
            pass
