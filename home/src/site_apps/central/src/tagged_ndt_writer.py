import openpyxl
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.styles import Alignment

def safe_write(ws, row, col, value):
    """병합 셀 충돌 없이 안전하게 값 쓰기 (앵커 셀에만 기입)"""
    for merge in ws.merged_cells.ranges:
        if merge.min_row <= row <= merge.max_row and merge.min_col <= col <= merge.max_col:
            # 앵커 셀이 아니면 무시 (에러 방지)
            if row != merge.min_row or col != merge.min_col:
                return
            ws.cell(row=merge.min_row, column=merge.min_col).value = value
            return
    ws.cell(row=row, column=col).value = value


def find_tag_position(ws, tag):
    """엑셀 시트에서 특정 텍스트 태그를 찾아 (row, col) 위치 반환 후 태그 삭제"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value.strip()
                tag_clean = tag.replace('[[', '').replace(']]', '')
                if tag_clean in v:
                    cell.value = None
                    return cell.row, cell.column
    return None, None


def insert_rows_safely(ws, insert_idx, amount):
    """openpyxl의 병합 셀 깨짐 버그를 우회하여 안전하게 행을 삽입합니다."""
    # 1. 모든 병합 정보 백업
    old_merges = [str(m) for m in ws.merged_cells.ranges]
    # 2. 병합 정보 초기화 (에러 원인 제거)
    ws.merged_cells.ranges = []
    
    # 3. 안전하게 빈 행 삽입
    ws.insert_rows(insert_idx, amount)
    
    # 4. 백업된 병합 정보를 위치에 맞게 수정하여 다시 적용
    for m_str in old_merges:
        cr = CellRange(m_str)
        if cr.min_row >= insert_idx:
            # 삽입 위치보다 아래에 있는 병합셀은 밀어냄
            cr.shift(row_shift=amount)
        elif cr.min_row < insert_idx <= cr.max_row:
            # 삽입 위치에 걸쳐있는 병합셀은 늘림
            cr.expand(down=amount)
        ws.merge_cells(str(cr))


def add_subtotals(records):
    """업체별 정렬 및 소계 행 추가"""
    if not records:
        return []
    
    sorted_recs = sorted(records, key=lambda x: str(x.get('업체', '')))
    new_records = []
    
    current_company = None
    sub_ori = 0.0
    sub_re = 0.0
    
    for rec in sorted_recs:
        comp = str(rec.get('업체', '')).strip()
        
        if current_company is not None and comp != current_company:
            # 이전 업체의 소계 추가
            new_records.append({
                '업체': f"[{current_company} 소계]",
                'is_subtotal': True,
                'ORI': sub_ori,
                'RE': sub_re
            })
            sub_ori = 0.0
            sub_re = 0.0
            
        current_company = comp
        new_records.append(rec)
        sub_ori += float(rec.get('ORI', 0) or 0)
        sub_re += float(rec.get('RE', 0) or 0)
        
    if current_company is not None:
        new_records.append({
            '업체': f"[{current_company} 소계]",
            'is_subtotal': True,
            'ORI': sub_ori,
            'RE': sub_re
        })
        
    return new_records


def write_tagged_records(ws, tag, records, col_map):
    """
    태그를 찾아 그 위치부터 records를 기입합니다.
    데이터 개수만큼 자동으로 행을 삽입하고, 첫 행의 병합 서식을 복사합니다.
    """
    if not records:
        return 0

    start_row, start_col = find_tag_position(ws, tag)
    if not start_row:
        return 0

    if len(records) > 1:
        insert_count = len(records) - 1
        insert_idx = start_row + 1
        
        # 첫 행(start_row)의 병합 정보를 미리 수집
        merges_to_copy = []
        for merge in ws.merged_cells.ranges:
            if merge.min_row == start_row and merge.max_row == start_row:
                merges_to_copy.append((merge.min_col, merge.max_col))
                
        # 안전한 행 삽입 (TOTAL 행 등 병합 깨짐 방지)
        insert_rows_safely(ws, insert_idx, insert_count)
        
        # 첫 행의 병합 정보를 새로 삽입된 빈 행들에 복사
        for r_offset in range(1, insert_count + 1):
            new_r = start_row + r_offset
            for min_col, max_col in merges_to_copy:
                try:
                    ws.merge_cells(start_row=new_r, start_column=min_col, end_row=new_r, end_column=max_col)
                except:
                    pass

    current_row = start_row
    item_no = 1
    
    from openpyxl.styles import Font
    bold_font = Font(bold=True)
    
    for rec in records:
        is_sub = rec.get('is_subtotal', False)
        
        if 'company' in col_map:
            safe_write(ws, current_row, col_map['company'], rec.get('업체', ''))
            if is_sub:
                ws.cell(row=current_row, column=col_map['company']).font = bold_font
                
        if not is_sub:
            if 'no' in col_map:
                safe_write(ws, current_row, col_map['no'], item_no)
                item_no += 1
            if 'section' in col_map:
                safe_write(ws, current_row, col_map['section'], rec.get('구간', ''))
            if 'line_no' in col_map:
                safe_write(ws, current_row, col_map['line_no'], rec.get('라인번호', ''))
            if 'pipe_size' in col_map:
                safe_write(ws, current_row, col_map['pipe_size'], rec.get('관경', ''))
            if 'joint' in col_map:
                safe_write(ws, current_row, col_map['joint'], rec.get('Joint No.', ''))
                
            if 'welder' in col_map:
                safe_write(ws, current_row, col_map['welder'], rec.get('용접사', ''))
            elif 'shift' in col_map:
                safe_write(ws, current_row, col_map['shift'], rec.get('규격', '주간'))
                
            if 'result' in col_map:
                safe_write(ws, current_row, col_map['result'], rec.get('결과', '합격'))
            elif 'unit' in col_map:
                safe_write(ws, current_row, col_map['unit'], 'M')

        ori = float(rec.get('ORI', 0) or 0)
        re_val = float(rec.get('RE', 0) or 0)
        tot = ori + re_val

        ws.row_dimensions[current_row].height = 45 # Line No. 높이 조절
        
        # Line No. 줄바꿈 처리
        if 'line_no' in col_map and not is_sub:
            # 병합된 경우 앵커 셀에 적용
            for merge in ws.merged_cells.ranges:
                if merge.min_row <= current_row <= merge.max_row and merge.min_col <= col_map['line_no'] <= merge.max_col:
                    ws.cell(merge.min_row, merge.min_col).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                    break
            else:
                ws.cell(current_row, col_map['line_no']).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

        # PAUT 길이 밑칸(ORI, RE, TOTAL) 병합 안전하게 해제
        target_cols = [col_map.get('ori'), col_map.get('re'), col_map.get('total')]
        target_cols = [c for c in target_cols if c is not None]
        
        new_ranges = []
        for merge in ws.merged_cells.ranges:
            if merge.min_row <= current_row <= merge.max_row and any(merge.min_col <= c <= merge.max_col for c in target_cols):
                continue # 제외 (병합 해제)
            new_ranges.append(merge)
        ws.merged_cells.ranges = new_ranges

        shrink_align = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        if 'ori' in col_map and ori > 0:
            safe_write(ws, current_row, col_map['ori'], round(ori, 4))
            ws.cell(row=current_row, column=col_map['ori']).alignment = shrink_align
            if is_sub: ws.cell(row=current_row, column=col_map['ori']).font = bold_font
        if 're' in col_map and re_val > 0:
            safe_write(ws, current_row, col_map['re'], round(re_val, 4))
            ws.cell(row=current_row, column=col_map['re']).alignment = shrink_align
            if is_sub: ws.cell(row=current_row, column=col_map['re']).font = bold_font
        if 'total' in col_map and tot > 0:
            safe_write(ws, current_row, col_map['total'], round(tot, 4))
            ws.cell(row=current_row, column=col_map['total']).alignment = shrink_align
            if is_sub: ws.cell(row=current_row, column=col_map['total']).font = bold_font

        current_row += 1

    return len(records)


def write_all_tagged_sections(ws, history, target_month_str, log_func=print):
    from site_apps.central.src.ndt_section_writer import extract_records_by_method
    
    paut_recs = extract_records_by_method(history, target_month_str, 'PAUT')
    paut_recs = add_subtotals(paut_recs)
    
    col_map_121 = {
        'company': 2, 'no': 3, 'section': 4, 'line_no': 6, 
        'pipe_size': 10, 'joint': 12, 'shift': 14, 'unit': 16, 
        'ori': 17, 're': 18, 'total': 20
    }
    w1 = write_tagged_records(ws, '[[NDT_121_PAUT]]', paut_recs, col_map_121)
    if w1:
        log_func(f"[OK] NDT_121_PAUT tag: {w1} records written")
        
    col_map_result = {
        'company': 2, 'no': 3, 'section': 4, 'line_no': 6, 
        'pipe_size': 10, 'joint': 12, 'welder': 14, 'result': 16, 
        'ori': 17, 're': 18, 'total': 20
    }
    w2 = write_tagged_records(ws, '[[NDT_RESULT_PAUT]]', paut_recs, col_map_result)
    if w2:
        log_func(f"[OK] NDT_RESULT_PAUT tag: {w2} records written")

    return True
