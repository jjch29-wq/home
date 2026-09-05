import openpyxl
import copy
import math
import re
from openpyxl.styles import Alignment
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.pagebreak import Break


def safe_write(ws, row, col, value):
    """병합 셀 충돌 없이 안전하게 값 쓰기 (앵커 셀에만 기입)"""
    for merge in ws.merged_cells.ranges:
        if merge.min_row <= row <= merge.max_row and merge.min_col <= col <= merge.max_col:
            ws.cell(row=merge.min_row, column=merge.min_col).value = value
            return
    ws.cell(row=row, column=col).value = value


def find_section_by_title(ws, title_keywords):
    """
    섹션 제목 키워드로 제목 행을 찾고, 그 다음 업체 헤더 행을 탐색합니다.
    Returns: (title_row, header_row, data_start_row)
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value.strip()
                if all(kw in v for kw in title_keywords):
                    title_row = cell.row
                    # 바로 아래에서 '업체' 헤더 찾기
                    for hr in range(title_row + 1, title_row + 10):
                        for col in range(1, 10):
                            hc = ws.cell(row=hr, column=col)
                            if hc.value and '업체' in str(hc.value):
                                return title_row, hr, hr + 1
    return None, None, None


def build_col_map_from_row(ws, header_row):
    """헤더 행을 읽어서 컬럼 매핑을 만듭니다."""
    col_map = {}
    for col in range(1, 30):
        c = ws.cell(row=header_row, column=col)
        if c.value and isinstance(c.value, str):
            v = c.value.strip()
            if '업체' in v:
                col_map['company'] = col
            elif '순번' in v or v in ('번호', 'No'):
                col_map['no'] = col
            elif 'Section' in v or '구간' == v:
                col_map['section'] = col
            elif 'Line No' in v or '라인번호' == v:
                col_map['line_no'] = col
            elif '관경' in v or '구경' in v:
                col_map['pipe_size'] = col
            elif '용접개소' in v or 'Joint' in v:
                col_map['joint'] = col
            elif '용접사' in v:
                col_map['welder'] = col
            elif '규격' == v:
                col_map['shift'] = col
            elif '단위' in v:
                col_map['unit'] = col
            elif '결과' in v:
                col_map['result'] = col
            elif 'PAUT 길이' in v or ('PAUT' in v and '길이' in v):
                col_map['ori'] = col
            elif '촬영매수' in v or '필름' in v:
                col_map['ori'] = col
            elif 'PT 길이' in v:
                col_map['ori'] = col
            elif 'MT 길이' in v:
                col_map['ori'] = col
            elif '비고' in v:
                col_map['note'] = col

    # 서브헤더 행(header_row+1)에서 ORI/RE/TOTAL 찾기
    for col in range(1, 30):
        c = ws.cell(row=header_row + 1, column=col)
        if c.value and isinstance(c.value, str):
            cv = c.value.strip().replace("'", "")
            if cv == 'ORI':
                col_map['ori'] = col
            elif cv == 'RE':
                col_map['re'] = col
            elif 'TOTAL' in cv:
                col_map['total'] = col

    return col_map


def write_records_to_section(ws, records, data_start, col_map):
    """
    records 리스트를 data_start 행부터 기입합니다.
    records: list of dict (업체, 구간, 라인번호, 관경, Joint No., 용접사, ORI, RE, 규격 등)
    """
    if not records:
        return 0

    current_row = data_start
    for i, rec in enumerate(records):
        if 'company' in col_map:
            safe_write(ws, current_row, col_map['company'], rec.get('업체', ''))
        if 'no' in col_map:
            safe_write(ws, current_row, col_map['no'], i + 1)
        if 'section' in col_map:
            safe_write(ws, current_row, col_map['section'], rec.get('구간', ''))
        if 'line_no' in col_map:
            safe_write(ws, current_row, col_map['line_no'], rec.get('라인번호', ''))
        if 'pipe_size' in col_map:
            safe_write(ws, current_row, col_map['pipe_size'], rec.get('관경', ''))
        if 'joint' in col_map:
            safe_write(ws, current_row, col_map['joint'], rec.get('Joint No.', ''))
        if 'welder' in col_map:
            safe_write(ws, current_row, col_map['welder'], rec.get('용접사', ''))
        if 'shift' in col_map:
            safe_write(ws, current_row, col_map['shift'], rec.get('규격', '주간'))
        if 'unit' in col_map:
            safe_write(ws, current_row, col_map['unit'], 'M')
        if 'result' in col_map:
            safe_write(ws, current_row, col_map['result'], rec.get('결과', '합격'))

        ori = float(rec.get('ORI', 0) or 0)
        re_val = float(rec.get('RE', 0) or 0)
        tot = ori + re_val

        ws.row_dimensions[current_row].height = 45 # Line No. 높이 조절
        
        if 'line_no' in col_map:
            for merge in ws.merged_cells.ranges:
                if merge.min_row <= current_row <= merge.max_row and merge.min_col <= col_map['line_no'] <= merge.max_col:
                    ws.cell(merge.min_row, merge.min_col).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                    break
            else:
                ws.cell(current_row, col_map['line_no']).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

        # PAUT 길이 밑칸(ORI, RE, TOTAL) 병합 안전하게 해제
        target_cols = [col_map.get('ori'), col_map.get('re'), col_map.get('total')]
        target_cols = [c for c in target_cols if c is not None]
        
        new_ranges = []
        for merge in ws.merged_cells.ranges:
            if merge.min_row <= current_row <= merge.max_row and any(merge.min_col <= c <= merge.max_col for c in target_cols):
                continue # 제외 (병합 해제)
            new_ranges.append(merge)
        ws.merged_cells.ranges = new_ranges

        shrink_align = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        if 'ori' in col_map and ori > 0:
            safe_write(ws, current_row, col_map['ori'], round(ori, 4))
            ws.cell(row=current_row, column=col_map['ori']).alignment = shrink_align
        if 're' in col_map and re_val > 0:
            safe_write(ws, current_row, col_map['re'], round(re_val, 4))
            ws.cell(row=current_row, column=col_map['re']).alignment = shrink_align
        if 'total' in col_map and tot > 0:
            safe_write(ws, current_row, col_map['total'], round(tot, 4))
            ws.cell(row=current_row, column=col_map['total']).alignment = shrink_align

        current_row += 1

    return len(records)


def _delete_rows_safely(ws, delete_idx, amount):
    """병합 범위를 보존하면서 행을 삭제한다."""
    if amount <= 0:
        return
    old_merges = [CellRange(str(m)) for m in ws.merged_cells.ranges]
    ws.merged_cells.ranges = []
    ws.delete_rows(delete_idx, amount)
    delete_end = delete_idx + amount - 1
    for cr in old_merges:
        if cr.max_row < delete_idx:
            ws.merge_cells(str(cr))
        elif cr.min_row > delete_end:
            cr.shift(row_shift=-amount)
            ws.merge_cells(str(cr))
        # 삭제 범위와 겹치는 병합은 빈 하단 행에만 사용되므로 제거한다.


def _insert_rows_safely(ws, insert_idx, amount):
    """병합 범위를 보존하면서 빈 행을 삽입한다."""
    if amount <= 0:
        return
    old_merges = [CellRange(str(m)) for m in ws.merged_cells.ranges]
    ws.merged_cells.ranges = []
    ws.insert_rows(insert_idx, amount)
    for cr in old_merges:
        if cr.min_row >= insert_idx:
            cr.shift(row_shift=amount)
        elif cr.min_row < insert_idx <= cr.max_row:
            cr.expand(down=amount)
        ws.merge_cells(str(cr))


def _copy_row_layout(ws, source_row, target_rows):
    """한 행의 셀 서식과 단일행 병합을 새 데이터 행들에 복사한다."""
    row_merges = [
        (m.min_col, m.max_col)
        for m in ws.merged_cells.ranges
        if m.min_row == source_row and m.max_row == source_row
    ]
    for target_row in target_rows:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
        for col in range(1, ws.max_column + 1):
            src = ws.cell(source_row, col)
            dst = ws.cell(target_row, col)
            if src.has_style:
                dst._style = copy.copy(src._style)
            dst.number_format = src.number_format
            dst.alignment = copy.copy(src.alignment)
            dst.protection = copy.copy(src.protection)
        for min_col, max_col in row_merges:
            try:
                ws.merge_cells(
                    start_row=target_row, start_column=min_col,
                    end_row=target_row, end_column=max_col
                )
            except ValueError:
                pass


def _clone_page_block(ws, page_start, page_end, insert_at):
    """수동 페이지 한 블록을 서식·병합·행높이와 함께 복제한다."""
    page_height = page_end - page_start + 1
    source_merges = [
        CellRange(str(m)) for m in ws.merged_cells.ranges
        if page_start <= m.min_row and m.max_row <= page_end
    ]
    old_break_ids = sorted(int(b.id) for b in ws.row_breaks.brk)

    # 병합을 안전하게 이동한 뒤 새 공간에 원본 페이지를 복사한다.
    old_merges = [CellRange(str(m)) for m in ws.merged_cells.ranges]
    ws.merged_cells.ranges = []
    ws.insert_rows(insert_at, page_height)
    for cr in old_merges:
        if cr.min_row >= insert_at:
            cr.shift(row_shift=page_height)
        elif cr.min_row < insert_at <= cr.max_row:
            cr.expand(down=page_height)
        ws.merge_cells(str(cr))

    offset = insert_at - page_start
    for source_row in range(page_start, page_end + 1):
        target_row = source_row + offset
        src_dim = ws.row_dimensions[source_row]
        dst_dim = ws.row_dimensions[target_row]
        dst_dim.height = src_dim.height
        dst_dim.hidden = src_dim.hidden
        dst_dim.outlineLevel = src_dim.outlineLevel
        for col in range(1, ws.max_column + 1):
            src = ws.cell(source_row, col)
            dst = ws.cell(target_row, col)
            dst.value = src.value
            if src.has_style:
                dst._style = copy.copy(src._style)
            dst.number_format = src.number_format
            dst.alignment = copy.copy(src.alignment)
            dst.protection = copy.copy(src.protection)

    for cr in source_merges:
        cr.shift(row_shift=offset)
        ws.merge_cells(str(cr))

    # 기존 이후 페이지 나눔은 아래로 이동하고 복제 페이지 끝에 새 나눔을 둔다.
    ws.row_breaks.brk = []
    for break_id in old_break_ids:
        shifted = break_id + page_height if break_id >= insert_at else break_id
        ws.row_breaks.append(Break(id=shifted, min=0, max=16383))
    ws.row_breaks.append(Break(id=insert_at + page_height - 1, min=0, max=16383))


def _page_bounds_for_row(ws, row):
    break_ids = sorted(int(b.id) for b in ws.row_breaks.brk)
    previous = 0
    for break_id in break_ids:
        if row <= break_id:
            return previous + 1, break_id
        previous = break_id
    return previous + 1, max(ws.max_row, row)


def write_rt_detail_with_page_overflow(ws, history, target_month_str, log_func=print):
    """
    2.2 MT와 2.4 PT 데이터가 없을 때 2.3 RT가 PT 아래 빈 행을 재사용한다.
    13페이지 용량을 넘으면 해당 페이지 블록 전체를 복제하여 RT를 이어 쓴다.
    """
    mt_records = extract_records_by_method(history, target_month_str, 'MT')
    rt_records = extract_records_by_method(history, target_month_str, 'RT')
    pt_records = extract_records_by_method(history, target_month_str, 'PT')
    if not rt_records or mt_records or pt_records:
        return False

    mt_title, _, _ = find_section_by_title(ws, ['2.2', 'MT'])
    rt_title, rt_header, rt_data_start = find_section_by_title(ws, ['2.3', 'RT'])
    pt_title, _, pt_data_start = find_section_by_title(ws, ['2.4', 'PT'])
    if not all((mt_title, rt_title, rt_header, rt_data_start, pt_title, pt_data_start)):
        return False

    col_map = build_col_map_from_row(ws, rt_header)
    if not col_map:
        return False

    page_start, page_end = _page_bounds_for_row(ws, rt_title)
    page_height = page_end - page_start + 1
    base_capacity = max(0, pt_title - rt_data_start)
    borrowed_capacity = max(0, page_end - pt_data_start + 1)
    page_capacity = base_capacity + borrowed_capacity
    if page_capacity <= 0:
        return False

    page_count = math.ceil(len(rt_records) / page_capacity)
    for page_index in range(1, page_count):
        insert_at = page_end + 1 + (page_index - 1) * page_height
        _clone_page_block(ws, page_start, page_end, insert_at)

    if page_count > 1 and ws.print_area:
        print_area_text = str(ws.print_area)
        area_match = re.search(
            r'\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)',
            print_area_text
        )
        if area_match:
            min_col, min_row, max_col, max_row = area_match.groups()
            extended_end = int(max_row) + page_height * (page_count - 1)
            ws.print_area = (
                f"{min_col}{min_row}:{max_col}{extended_end}"
            )

    template_row_height = ws.row_dimensions[rt_data_start].height
    written = 0
    for page_index in range(page_count):
        offset = page_index * page_height
        chunk = rt_records[written:written + page_capacity]
        chunk_rt_start = rt_data_start + offset
        chunk_pt_title = pt_title + offset
        chunk_page_end = page_end + offset
        extra_rows = max(0, len(chunk) - base_capacity)
        if extra_rows:
            # PT 제목을 아래로 밀어 RT 행을 만들고, 페이지 하단 빈 행을 같은 수만큼 제거한다.
            _insert_rows_safely(ws, chunk_pt_title, extra_rows)
            _copy_row_layout(
                ws, chunk_rt_start,
                range(chunk_pt_title, chunk_pt_title + extra_rows)
            )
            _delete_rows_safely(ws, chunk_page_end + 1, extra_rows)

        write_records_to_section(
            ws, chunk, chunk_rt_start, col_map,
            row_height=template_row_height
        )
        written += len(chunk)

    log_func(
        f"✅ 2.3 RT: {written}건 기입, PT 하단 빈 행 재사용, "
        f"13페이지 복제 {page_count - 1}회"
    )
    return True


def extract_records_by_method(history, target_month_str, method_filter):
    """
    history에서 특정 검사방법(PAUT/RT/MT/PT)의 레코드를 추출합니다.
    ORI/RE를 분류하여 list of dict으로 반환합니다.
    """
    raw = []
    for date_key, log_data in history.items():
        if date_key.startswith(target_month_str):
            for r in log_data.get('ndt_results', []):
                method = str(r.get('검사방법', '')).strip().upper()
                if method == method_filter.upper():
                    r['_date'] = date_key
                    raw.append(r)

    raw.sort(key=lambda x: x['_date'])

    groups = {}
    for r in raw:
        key = (
            str(r.get('업체', '')),
            str(r.get('구간', '')),
            str(r.get('라인번호', '')),
            str(r.get('관경', '')),
            str(r.get('Joint No.', '')),
            str(r.get('용접사', ''))
        )
        # 해당 방법의 수치 필드 찾기
        method = method_filter.upper()
        val_fields = {
            'PAUT': ('PAUT',),
            'RT': ('RT_OR', 'RT_ORI', 'RT'),
            'MT': ('MT',),
            'PT': ('PT',),
        }
        val = 0.0
        for field in val_fields.get(method, (method,)):
            raw_val = str(r.get(field, '') or '').strip()
            if raw_val:
                try:
                    v = float(raw_val)
                    if v > 0:
                        val = v
                        break
                except:
                    pass

        if val == 0:
            continue

        shift = str(r.get('규격', '주간')).strip()
        if key not in groups:
            groups[key] = {'ORI': 0.0, 'RE': 0.0, 'shift': shift}
        if groups[key]['ORI'] == 0.0:
            groups[key]['ORI'] += val
        else:
            groups[key]['RE'] += val

    result_list = []
    for (업체, 구간, 라인번호, 관경, joint, 용접사), data in groups.items():
        result_list.append({
            '업체': 업체, '구간': 구간, '라인번호': 라인번호,
            '관경': 관경, 'Joint No.': joint, '용접사': 용접사,
            '규격': data['shift'], '결과': '합격',
            'ORI': data['ORI'], 'RE': data['RE']
        })
    return result_list


def write_all_ndt_sections(ws, history, target_month_str, log_func=print):
    """
    엑셀 시트에서 모든 NDT 섹션(1.2.2~1.2.4, 2.1~2.4)을 찾아 데이터를 기입합니다.
    """
    # 섹션 정의: (title_keywords, method_filter, section_label)
    sections = [
        (['1.2.1', 'PAUT'],      'PAUT', '1.2.1 PAUT'),
        (['1.2.2', 'MT'],        'MT',   '1.2.2 MT'),
        (['1.2.2', '자분'],      'MT',   '1.2.2 자분(MT)'),
        (['1.2.3', 'RT'],        'RT',   '1.2.3 RT'),
        (['1.2.3', '방사선'],    'RT',   '1.2.3 방사선(RT)'),
        (['1.2.4', 'PT'],        'PT',   '1.2.4 PT'),
        (['1.2.4', '침투'],      'PT',   '1.2.4 침투(PT)'),
        (['2.1', 'PAUT'],        'PAUT', '2.1 PAUT'),
        (['2.2', 'MT'],          'MT',   '2.2 MT'),
        (['2.3', 'RT'],          'RT',   '2.3 RT'),
        (['2.4', 'PT'],          'PT',   '2.4 PT'),
    ]

    written_sections = set()
    for kws, method, label in sections:
        if label in written_sections:
            continue
        title_row, header_row, data_start = find_section_by_title(ws, kws)
        if not header_row:
            continue

        col_map = build_col_map_from_row(ws, header_row)
        if not col_map:
            log_func(f"⚠️ {label}: 헤더 컬럼 매핑 실패")
            continue

        records = extract_records_by_method(history, target_month_str, method)
        if not records:
            log_func(f"ℹ️ {label}: {method} 데이터 없음")
            written_sections.add(label)
            continue

        written = write_records_to_section(ws, records, data_start, col_map)
        log_func(f"✅ {label}: {written}건 기입 완료 (row {data_start}~, col_map={col_map})")
        written_sections.add(label)

    return True


# ====================
# 테스트 실행
# ====================
if __name__ == '__main__':
    import json, sys
    sys.path.insert(0, r'c:\Users\jjch2\Desktop\PMI\home\src')
    from site_apps.central.src.paut_writer import find_paut_section, write_paut_data

    history_path = r'c:\Users\jjch2\Desktop\PMI\home\src\daily_work_history.json'
    with open(history_path, 'r', encoding='utf-8') as f:
        history = json.load(f)

    wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\템플릿_최종완성본_V74.xlsx')
    ws = wb.worksheets[0]

    target_month = '2026-08'
    write_all_ndt_sections(ws, history, target_month)

    wb.save(r'C:\Users\jjch2\Desktop\Test_AllNDT.xlsx')
    print("Done! Saved to C:\\Users\\jjch2\\Desktop\\Test_AllNDT.xlsx")
