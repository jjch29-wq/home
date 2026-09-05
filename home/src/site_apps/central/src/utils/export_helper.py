import os
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def export_budget_estimation_to_excel(filepath, summary_data, labor_data, material_data, expense_data):
    """
    Export the detailed budget estimation to a formatted Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "공사실행예산서"
    
    # Styles
    title_font = Font(name='맑은 고딕', size=16, bold=True)
    header_font = Font(name='맑은 고딕', size=11, bold=True)
    bold_font = Font(name='맑은 고딕', size=10, bold=True)
    normal_font = Font(name='맑은 고딕', size=10)
    
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    thin = Side(border_style="thin", color="000000")
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    sub_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    blue_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    def to_float(val):
        try:
            return float(str(val).replace(',', '').replace('₩', '').replace('%', '').strip() or 0)
        except:
            return 0.0

    # 1. Title
    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = f"[{summary_data.get('site', '')}] 공사실행예산서 (사정원가)"
    c.font = Font(name='맑은 고딕', size=18, bold=True)
    c.alignment = center_align
    ws.row_dimensions[1].height = 30
    
    # Write metadata (Print Date)
    ws.merge_cells('A2:B2')
    ws['A2'] = "출력일시: " + datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws['A2'].font = normal_font
    ws['A2'].alignment = Alignment(horizontal='left', vertical='top')

    # 1.5 Approval Box (결재란)
    approval_headers = ["소장", "담당", "상무", "본부장", "사장"]
    start_col = 3 # Column C
    
    for i, h in enumerate(approval_headers):
        col_letter = get_column_letter(start_col + i)
        
        # Header cell (Row 2)
        h_cell = ws[f"{col_letter}2"]
        h_cell.value = h
        h_cell.font = bold_font
        h_cell.alignment = center_align
        h_cell.border = border_all
        
        # Stamp cell (Row 3)
        s_cell = ws[f"{col_letter}3"]
        s_cell.border = border_all
        s_cell.alignment = center_align
        
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 60
    
    # Helper function to write a table
    def write_table(start_row, title, headers, data_rows, total_label=None, total_val=None):
        # Table Title
        ws.cell(row=start_row, column=1, value=title).font = header_font
        ws.cell(row=start_row, column=1).fill = header_fill
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
        for col in range(1, len(headers)+1):
            ws.cell(row=start_row, column=col).border = border_all
            ws.cell(row=start_row, column=col).alignment = center_align
            
        # Headers
        r = start_row + 1
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = bold_font
            c.fill = sub_fill
            c.border = border_all
            c.alignment = center_align
                
        # Data Rows
        r += 1
        for row_data in data_rows:
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.font = normal_font
                c.border = border_all
                if isinstance(val, (int, float)) or (isinstance(val, str) and val.replace(',', '').replace('.', '').replace('-', '').isdigit()):
                    c.number_format = '#,##0'
                    c.alignment = right_align
                    if isinstance(val, str):
                        try: c.value = float(val.replace(',', ''))
                        except: pass
                else:
                    c.alignment = center_align
            r += 1
            
        # Total Row
        if total_label is not None:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers)-1)
            c = ws.cell(row=r, column=1, value=total_label)
            c.font = bold_font
            c.fill = yellow_fill
            c.border = border_all
            c.alignment = center_align
            for col in range(2, len(headers)):
                ws.cell(row=r, column=col).border = border_all
                
            c_tot = ws.cell(row=r, column=len(headers), value=total_val)
            c_tot.font = bold_font
            c_tot.fill = yellow_fill
            c_tot.border = border_all
            c_tot.alignment = right_align
            c_tot.number_format = '#,##0'
            try: c_tot.value = float(str(total_val).replace(',', '').replace('₩', '').strip())
            except: pass
            r += 1
            
        return r + 1

    # Set Global Column Widths to ensure all text and large numbers fit perfectly
    # Set Global Column Widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12

    # Print / Page Setup: Fit to one page wide, Landscape, A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 9


    # 1.8 Site Metadata Info
    site_info_row = 5
    info_font = Font(name='맑은 고딕', size=11, bold=True)
    
    ws.cell(row=site_info_row, column=1, value="현 장 명 :").font = info_font
    ws.cell(row=site_info_row, column=2, value=f"{summary_data.get('site', '')} 중 비파괴검사").font = info_font
    
    ws.cell(row=site_info_row+1, column=1, value="공사기간 :").font = info_font
    period_str = summary_data.get('period', '')
    ws.cell(row=site_info_row+1, column=2, value=period_str).font = info_font
    
    days_diff_str = ""
    if period_str and "~" in period_str:
        try:
            start_str, end_str = period_str.split("~")
            start_date = datetime.datetime.strptime(start_str.strip(), "%Y.%m.%d")
            end_date = datetime.datetime.strptime(end_str.strip(), "%Y.%m.%d")
            days_diff = (end_date - start_date).days + 1
            days_diff_str = f"{days_diff}"
        except Exception:
            pass

    ws.cell(row=site_info_row+1, column=5, value=days_diff_str).font = info_font
    ws.cell(row=site_info_row+1, column=5).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=site_info_row+1, column=6, value="일").font = info_font
    ws.cell(row=site_info_row+1, column=6).alignment = Alignment(horizontal='left', vertical='center')
    
    ws.cell(row=site_info_row+2, column=1, value="계약업체 :").font = info_font
    ws.cell(row=site_info_row+2, column=2, value="").font = info_font
    
    ws.cell(row=site_info_row+3, column=1, value="경인사무소 :").font = info_font
    ws.cell(row=site_info_row+3, column=2, value="").font = info_font

    current_row = site_info_row + 5

    # 2. Summary Table
    ws.cell(row=current_row, column=1, value="1) 사전예산 요약").font = header_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    for col in range(1, 3):
        ws.cell(row=current_row, column=col).border = border_all
        ws.cell(row=current_row, column=col).alignment = center_align
    current_row += 1

    summary_items = [
        ("계약금액(Revenue)", summary_data.get('revenue', 0)),
        ("매출금액(UnitPrice)", summary_data.get('unitprice', 0)),
        ("실행 노무비(Labor)", summary_data.get('labor', 0)),
        ("실행 재료비(Material)", summary_data.get('material', 0)),
        ("실행 경비(Expense)", summary_data.get('expense', 0)),
        ("실행 외주비(Outsource)", summary_data.get('outsource', 0)),
        ("영업이익(Profit)", summary_data.get('profit', 0)),
        ("이익률(Margin)", summary_data.get('margin', '0%'))
    ]
    
    for k, v in summary_items:
        ws.cell(row=current_row, column=1, value=k).font = bold_font
        ws.cell(row=current_row, column=1).fill = sub_fill
        ws.cell(row=current_row, column=1).border = border_all
        
        if k == "이익률(Margin)" and '%' not in str(v):
            v = f"{v}%"
            
        c_val = ws.cell(row=current_row, column=2, value=v)
        c_val.font = normal_font
        c_val.border = border_all
        if '%' in str(v):
            c_val.alignment = right_align
        else:
            c_val.number_format = '#,##0'
            c_val.alignment = right_align
            try: c_val.value = float(str(v).replace(',', '').replace('₩', '').strip())
            except: pass
        current_row += 1
        
    current_row += 2

    # 3. Labor Data
    if labor_data:
        # Table 1: Regular
        reg_rows = []
        reg_total = 0
        regular_keys = ['부장', '차장', '과장', '대리', '사원', '초급', '중급', '고급', '특급']
        for rank in regular_keys:
            if rank in labor_data:
                vals = labor_data[rank]
                p = to_float(vals.get('personnel'))
                d = to_float(vals.get('period'))
                u = to_float(vals.get('unit_price'))
                amt = p * d * u
                if p > 0 or d > 0:
                    reg_rows.append([rank, p, d, u, amt])
                    reg_total += amt
        
        if reg_rows:
            current_row = write_table(
                current_row, "2) 노무비 - 정시근무", 
                ["직급", "투입인원(명)", "투입기간(일)", "단가", "사전원가가액"], 
                reg_rows, "정시근무 합계", reg_total
            )

        # Table 2: Special
        sp_rows = []
        sp_total = 0
        special_keys = ['연장근무', '야간근무', '휴일근무']
        for stype in special_keys:
            if stype in labor_data:
                vals = labor_data[stype]
                p = to_float(vals.get('personnel'))
                h = to_float(vals.get('period'))
                u = to_float(vals.get('unit_price'))
                amt = p * h * u
                if p > 0 or h > 0:
                    sp_rows.append([stype, p, h, u, amt])
                    sp_total += amt
                
        if sp_rows:
            current_row = write_table(
                current_row, "2) 노무비 - 특별근무", 
                ["형태", "투입인원(명)", "투입시간(시간)", "단가", "사전원가가액"], 
                sp_rows, "특별근무 합계", sp_total
            )

    # 4. Material Data
    if material_data:
        mat_rows = []
        mat_total = 0
        for item in material_data:
            qty = to_float(item.get('qty'))
            price = to_float(item.get('price'))
            amt = qty * price
            if qty > 0:
                mat_rows.append([item.get('name'), item.get('spec'), qty, item.get('unit'), price, amt])
                mat_total += amt
        if mat_rows:
            current_row = write_table(
                current_row, "3) 재료비", 
                ["품목", "사양", "수량", "규격", "단가", "사전원가가액"], 
                mat_rows, "재료비 합계", mat_total
            )
            
    # 5. Expense Data
    if expense_data:
        # Site Expense
        exp_rows = []
        exp_total = 0
        for item in expense_data.get('site_expense', []):
            cont = item.get('cont', '')
            ppl = to_float(item.get('ppl'))
            qty = to_float(item.get('qty'))
            price = to_float(item.get('price'))
            amt = qty * price
            if qty > 0 or ppl > 0:
                exp_rows.append([cont, ppl, qty, price, amt])
                exp_total += amt
        if exp_rows:
            current_row = write_table(
                current_row, "4) 경비 - 현장경비", 
                ["항목(내용)", "투입인원", "수량/일수", "단가", "사전원가가액"], 
                exp_rows, "현장경비 합계", exp_total
            )
            
        # Rental
        rent_rows = []
        rent_total = 0
        for item in expense_data.get('rental', []):
            spec = item.get('spec', '')
            qty = to_float(item.get('qty'))
            period = to_float(item.get('period'))
            price = to_float(item.get('price'))
            amt = qty * period * price
            if qty > 0 or period > 0:
                rent_rows.append([spec, qty, period, price, amt])
                rent_total += amt
        if rent_rows:
            current_row = write_table(
                current_row, "4) 경비 - 렌탈비", 
                ["사양", "수량", "기간(일수)", "단가", "사전원가가액"], 
                rent_rows, "렌탈비 합계", rent_total
            )
            
        # Outsource
        out_rows = []
        out_total = 0
        for item in expense_data.get('outsource', []):
            work = item.get('work', '')
            count = to_float(item.get('count'))
            price = to_float(item.get('price'))
            amt = count * price
            if count > 0:
                out_rows.append([work, count, "", price, amt])
                out_total += amt
        if out_rows:
            current_row = write_table(
                current_row, "4) 경비 - 외주비", 
                ["작업내용", "인원/횟수", "비고", "단가", "사전원가가액"], 
                out_rows, "외주비 합계", out_total
            )
            
        # Depreciation
        dep_rows = []
        dep_total = 0
        for item in expense_data.get('depreciation', []):
            name = item.get('item', '')
            qty = to_float(item.get('qty'))
            days = to_float(item.get('days'))
            rate = to_float(item.get('rate'))
            amt = qty * days * rate
            if qty > 0 or days > 0:
                dep_rows.append([name, qty, days, rate, amt])
                dep_total += amt
        if dep_rows:
            current_row = write_table(
                current_row, "4) 경비 - 감가상각", 
                ["장비명", "대수", "사용일수", "상각율", "사전원가가액"], 
                dep_rows, "감가상각 합계", dep_total
            )

    wb.save(filepath)
    return True
