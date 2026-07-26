import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os
import json
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime

class WorkApprovalApp:
    def __init__(self, root):
        self.root = root
        self.root.title("작업승인계획서 자동 생성기")
        self.root.geometry("700x950")
        self.root.resizable(True, True)
        
        style = ttk.Style()
        style.theme_use('clam')
        
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        main_frame = ttk.Frame(self.notebook, padding=15)
        self.notebook.add(main_frame, text="작업승인계획서")
        
        # Add TBM Form Tab
        try:
            from tbm_tab import TBMFormTab
            self.tab_tbm = ttk.Frame(self.notebook)
            self.notebook.add(self.tab_tbm, text='TBM 회의록')
            self.tbm_manager = TBMFormTab(self.tab_tbm, main_app=self)
            self.tbm_manager.pack(fill='both', expand=True)
        except Exception as e:
            print(f"TBM 모듈 로드 실패: {e}")
            
        # Add Risk Assessment Tab
        try:
            import importlib.util
            import sys
            
            # 위험성 평가표 all.py 파일이 띄어쓰기가 있어 importlib 사용
            module_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "위험성 평가표 all.py")
            spec = importlib.util.spec_from_file_location("risk_module", module_path)
            risk_module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(risk_module)
            
            self.tab_risk = ttk.Frame(self.notebook)
            self.notebook.add(self.tab_risk, text='위험성 평가표')
            self.risk_manager = risk_module.RiskAssessmentApp(self.tab_risk)
        except Exception as e:
            print(f"위험성 평가표 모듈 로드 실패: {e}")
        
        # Title
        ttk.Label(main_frame, text="[서식 3] 작업승인계획서 생성기", font=('Malgun Gothic', 16, 'bold')).pack(pady=(0, 15))
        
        # Form Container using Canvas for scrolling
        bg_color = style.lookup('TFrame', 'background')
        canvas = tk.Canvas(main_frame, bg=bg_color, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        form_frame = ttk.Frame(canvas)

        canvas_window = canvas.create_window((0, 0), window=form_frame, anchor="nw")

        form_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )
        
        def on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)
            
        canvas.bind("<Configure>", on_canvas_configure)
        
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        row_idx = 0
        
        # 0. 날짜 및 기능
        date_frame = ttk.Frame(form_frame)
        date_frame.grid(row=row_idx, column=0, columnspan=2, sticky='w', pady=(0, 10))
        
        ttk.Label(date_frame, text="일자:").pack(side='left', padx=(5, 5))
        self.ent_date = ttk.Combobox(date_frame, width=25)
        self.ent_date.insert(0, datetime.now().strftime("%Y년 %m월 %d일 O요일"))
        self.ent_date.pack(side='left')
        self.ent_date.bind('<<ComboboxSelected>>', self.on_date_select)
        
        self.btn_generate = ttk.Button(date_frame, text="엑셀 생성", command=self.generate_files)
        self.btn_generate.pack(side='left', padx=(20, 5))
        
        self.btn_unified = ttk.Button(date_frame, text="🔥 통합 엑셀", command=self.generate_unified_excel)
        self.btn_unified.pack(side='left', padx=5)
        
        self.lbl_status = ttk.Label(date_frame, text="대기 중...", foreground="gray")
        self.lbl_status.pack(side='left', padx=10)
        
        row_idx += 1
        
        # 1. 작업 구분
        ttk.Label(form_frame, text="작업 구분:", font=('Malgun Gothic', 10, 'bold')).grid(row=row_idx, column=0, sticky='e', padx=5, pady=5)
        type_frame = ttk.Frame(form_frame)
        type_frame.grid(row=row_idx, column=1, sticky='w')
        self.var_work_type = tk.StringVar(value="위험")
        ttk.Radiobutton(type_frame, text="위험", variable=self.var_work_type, value="위험").pack(side='left', padx=10)
        ttk.Radiobutton(type_frame, text="일반", variable=self.var_work_type, value="일반").pack(side='left', padx=10)
        row_idx += 1
        
        # 2. 위험작업종류 (체크박스)
        ttk.Label(form_frame, text="위험작업종류:").grid(row=row_idx, column=0, sticky='ne', padx=5, pady=5)
        risk_types_frame = ttk.Frame(form_frame)
        risk_types_frame.grid(row=row_idx, column=1, sticky='w', pady=5)
        
        self.risk_vars = {}
        risk_list = ["화기작업", "밀폐공간 작업", "고소작업", "전기통신 작업", "화학물질 취급작업", 
                     "굴착작업", "중량물 취급작업", "지하공간 가스방출", "지붕위·철골작업", "기타 위험작업"]
        
        for i, r_type in enumerate(risk_list):
            var = tk.BooleanVar(value=False)
            self.risk_vars[r_type] = var
            cb = ttk.Checkbutton(risk_types_frame, text=r_type, variable=var)
            cb.grid(row=i//3, column=i%3, sticky='w', padx=5, pady=2)
            
        row_idx += 1
        ttk.Separator(form_frame, orient='horizontal').grid(row=row_idx, column=0, columnspan=2, sticky='ew', pady=10)
        row_idx += 1

        # 3. 작업 기본 정보
        ttk.Label(form_frame, text="작업명:").grid(row=row_idx, column=0, sticky='e', padx=5, pady=2)
        self.ent_work_name = ttk.Entry(form_frame, width=50)
        self.ent_work_name.grid(row=row_idx, column=1, sticky='w', padx=5, pady=2)
        row_idx += 1
        
        ttk.Label(form_frame, text="작업기간:").grid(row=row_idx, column=0, sticky='e', padx=5, pady=2)
        self.ent_work_period = ttk.Entry(form_frame, width=50)
        self.ent_work_period.grid(row=row_idx, column=1, sticky='w', padx=5, pady=2)
        row_idx += 1

        ttk.Label(form_frame, text="원도급 업체명:").grid(row=row_idx, column=0, sticky='e', padx=5, pady=2)
        self.ent_main_company = ttk.Entry(form_frame, width=50)
        self.ent_main_company.grid(row=row_idx, column=1, sticky='w', padx=5, pady=2)
        row_idx += 1
        
        ttk.Label(form_frame, text="원도급 작업책임자:").grid(row=row_idx, column=0, sticky='e', padx=5, pady=2)
        self.ent_main_manager = ttk.Entry(form_frame, width=50)
        self.ent_main_manager.grid(row=row_idx, column=1, sticky='w', padx=5, pady=2)
        row_idx += 1
        
        ttk.Label(form_frame, text="하도급 업체명:").grid(row=row_idx, column=0, sticky='e', padx=5, pady=2)
        self.ent_sub_company = ttk.Entry(form_frame, width=50)
        self.ent_sub_company.grid(row=row_idx, column=1, sticky='w', padx=5, pady=2)
        row_idx += 1
        
        ttk.Label(form_frame, text="하도급 작업책임자:").grid(row=row_idx, column=0, sticky='e', padx=5, pady=2)
        self.ent_sub_manager = ttk.Entry(form_frame, width=50)
        self.ent_sub_manager.grid(row=row_idx, column=1, sticky='w', padx=5, pady=2)
        row_idx += 1
        
        ttk.Separator(form_frame, orient='horizontal').grid(row=row_idx, column=0, columnspan=2, sticky='ew', pady=10)
        row_idx += 1

        # 4. 투입 현황
        ttk.Label(form_frame, text="작업인원:").grid(row=row_idx, column=0, sticky='ne', padx=5, pady=2)
        self.txt_personnel = tk.Text(form_frame, width=50, height=3, font=('Malgun Gothic', 9))
        self.txt_personnel.insert('1.0', "총원 명\n(직종별 인원 세부적으로 작성)")
        self.txt_personnel.grid(row=row_idx, column=1, sticky='w', padx=5, pady=2)
        row_idx += 1
        
        ttk.Label(form_frame, text="사용장비:").grid(row=row_idx, column=0, sticky='e', padx=5, pady=2)
        self.ent_equip = ttk.Entry(form_frame, width=50)
        self.ent_equip.insert(0, "(종류, 수량 기재)")
        self.ent_equip.grid(row=row_idx, column=1, sticky='w', padx=5, pady=2)
        row_idx += 1
        
        ttk.Label(form_frame, text="개인보호장비:").grid(row=row_idx, column=0, sticky='e', padx=5, pady=2)
        self.ent_ppe = ttk.Entry(form_frame, width=50)
        self.ent_ppe.insert(0, "(종류, 수량 기재)")
        self.ent_ppe.grid(row=row_idx, column=1, sticky='w', padx=5, pady=2)
        row_idx += 1
        
        ttk.Label(form_frame, text="기타 안전장비:").grid(row=row_idx, column=0, sticky='e', padx=5, pady=2)
        self.ent_safety_equip = ttk.Entry(form_frame, width=50)
        self.ent_safety_equip.insert(0, "(종류, 수량 기재)")
        self.ent_safety_equip.grid(row=row_idx, column=1, sticky='w', padx=5, pady=2)
        row_idx += 1
        
        ttk.Separator(form_frame, orient='horizontal').grid(row=row_idx, column=0, columnspan=2, sticky='ew', pady=10)
        row_idx += 1

        # 5. 상세 내용
        ttk.Label(form_frame, text="작업내용:").grid(row=row_idx, column=0, sticky='ne', padx=5, pady=2)
        self.txt_work_content = tk.Text(form_frame, width=50, height=5, font=('Malgun Gothic', 9))
        self.txt_work_content.grid(row=row_idx, column=1, sticky='w', padx=5, pady=2)
        row_idx += 1

        ttk.Label(form_frame, text="주요 위험요인:").grid(row=row_idx, column=0, sticky='ne', padx=5, pady=2)
        self.txt_risk_factors = tk.Text(form_frame, width=50, height=4, font=('Malgun Gothic', 9))
        self.txt_risk_factors.grid(row=row_idx, column=1, sticky='w', padx=5, pady=2)
        row_idx += 1
        
        ttk.Label(form_frame, text="현장점검\n체크리스트:").grid(row=row_idx, column=0, sticky='ne', padx=5, pady=2)
        self.txt_checklist = tk.Text(form_frame, width=50, height=3, font=('Malgun Gothic', 9))
        self.txt_checklist.insert('1.0', "전기작업, 굴착작업, 화기작업, 밀폐공간 작업시 작업책임자가 현장에서 체크리스트 점검사항 직접 확인 후 기입")
        self.txt_checklist.grid(row=row_idx, column=1, sticky='w', padx=5, pady=2)
        row_idx += 1
        
        # Load saved config
        self.load_config()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def get_history_path(self):
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'work_approval_history_v2.json')

    def load_history_data(self):
        path = self.get_history_path()
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        return {}

    def save_history_data(self, history):
        try:
            with open(self.get_history_path(), 'w', encoding='utf-8') as f:
                json.dump(history, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Failed to save history: {e}")

    def on_date_select(self, event=None):
        date_str = self.ent_date.get()
        history = self.load_history_data()
        if date_str in history:
            self.populate_ui(history[date_str], date_str)

    def save_config(self):
        history = self.load_history_data()
        current_date = self.ent_date.get()
        
        risk_selections = {k: v.get() for k, v in self.risk_vars.items()}
        
        data = {
            'work_type': self.var_work_type.get(),
            'risks': risk_selections,
            'work_name': self.ent_work_name.get(),
            'work_period': self.ent_work_period.get(),
            'main_company': self.ent_main_company.get(),
            'main_manager': self.ent_main_manager.get(),
            'sub_company': self.ent_sub_company.get(),
            'sub_manager': self.ent_sub_manager.get(),
            'personnel': self.txt_personnel.get('1.0', tk.END).strip(),
            'equip': self.ent_equip.get(),
            'ppe': self.ent_ppe.get(),
            'safety_equip': self.ent_safety_equip.get(),
            'work_content': self.txt_work_content.get('1.0', tk.END).strip(),
            'risk_factors': self.txt_risk_factors.get('1.0', tk.END).strip(),
            'checklist': self.txt_checklist.get('1.0', tk.END).strip()
        }
        
        history[current_date] = data
        self.save_history_data(history)
        
        dates = [k for k in history.keys() if not k.startswith('_')]
        self.ent_date['values'] = sorted(dates, reverse=True)

    def on_close(self):
        history = self.load_history_data()
        history['_window_geometry'] = self.root.geometry()
        self.save_history_data(history)
        self.root.destroy()

    def load_config(self):
        history = self.load_history_data()
        if not history:
            return

        if '_window_geometry' in history:
            pass # Keep logic clear of window geometry issues

        dates = [k for k in history.keys() if not k.startswith('_')]
        if not dates:
            return
            
        sorted_dates = sorted(dates, reverse=True)
        self.ent_date['values'] = sorted_dates

        most_recent_date = sorted_dates[0]
        self.ent_date.delete(0, tk.END)
        self.ent_date.insert(0, most_recent_date)
        
        self.populate_ui(history[most_recent_date], most_recent_date)

    def populate_ui(self, data, date_str=None):
        def set_ent(ent, val):
            if val is not None:
                ent.delete(0, tk.END)
                ent.insert(0, str(val))
                
        def set_txt(txt, val):
            if val is not None:
                txt.delete('1.0', tk.END)
                txt.insert('1.0', str(val))
        
        if date_str:
            set_ent(self.ent_date, date_str)
            
        if 'work_type' in data:
            self.var_work_type.set(data['work_type'])
            
        if 'risks' in data:
            for k, v in data['risks'].items():
                if k in self.risk_vars:
                    self.risk_vars[k].set(v)
                    
        set_ent(self.ent_work_name, data.get('work_name'))
        set_ent(self.ent_work_period, data.get('work_period'))
        set_ent(self.ent_main_company, data.get('main_company'))
        set_ent(self.ent_main_manager, data.get('main_manager'))
        set_ent(self.ent_sub_company, data.get('sub_company'))
        set_ent(self.ent_sub_manager, data.get('sub_manager'))
        set_txt(self.txt_personnel, data.get('personnel'))
        set_ent(self.ent_equip, data.get('equip'))
        set_ent(self.ent_ppe, data.get('ppe'))
        set_ent(self.ent_safety_equip, data.get('safety_equip'))
        set_txt(self.txt_work_content, data.get('work_content'))
        set_txt(self.txt_risk_factors, data.get('risk_factors'))
        set_txt(self.txt_checklist, data.get('checklist'))

    def generate_files(self, silent_path=None):
        self.save_config()
        current_date = self.ent_date.get()
        history = self.load_history_data()
        params = history.get(current_date, {})
        
        if silent_path:
            output_path = silent_path
        else:
            initial_dir = history.get('_last_save_dir', os.path.dirname(os.path.abspath(__file__)))
            output_dir = filedialog.askdirectory(title="저장할 폴더를 선택하세요", initialdir=initial_dir)
            if not output_dir:
                return
            history['_last_save_dir'] = output_dir
            self.save_history_data(history)
            output_path = os.path.join(output_dir, "작업승인계획서.xlsx")
        
        self.btn_generate.config(state='disabled')
        self.lbl_status.config(text="엑셀 파일 생성 중...", foreground="blue")
        self.root.update()
        
        try:
            self.create_excel(output_path, params)
            if not silent_path:
                messagebox.showinfo("생성 완료", f"작업승인계획서가 성공적으로 생성되었습니다!\n\n저장 위치:\n{output_path}")
                self.lbl_status.config(text="완료!", foreground="green")
            else:
                self.lbl_status.config(text="임시 파일 생성 완료!", foreground="green")
        except Exception as e:
            if not silent_path:
                messagebox.showerror("오류", f"엑셀 파일 생성 중 오류가 발생했습니다:\n{e}")
            self.lbl_status.config(text="오류 발생", foreground="red")
        finally:
            self.btn_generate.config(state='normal')

    def generate_unified_excel(self):
        history = self.load_history_data()
        initial_dir = history.get('_last_save_dir', os.path.dirname(os.path.abspath(__file__)))
        
        output_dir = filedialog.askdirectory(title="일일 안전서류 통합 엑셀 저장 폴더를 선택하세요", initialdir=initial_dir)
        if not output_dir: return
        
        history['_last_save_dir'] = output_dir
        self.save_history_data(history)
        
        date_str = self.ent_date.get().replace("-", "")
        final_path = os.path.join(output_dir, f"일일_안전서류_통합_{date_str}.xlsx").replace("/", "\\")
        
        self.lbl_status.config(text="통합 엑셀 생성 중...", foreground="blue")
        self.btn_generate.config(state='disabled')
        self.btn_unified.config(state='disabled')
        self.root.update()
        
        temp_approval = os.path.join(output_dir, f"temp_approval_{date_str}.xlsx").replace("/", "\\")
        temp_tbm = os.path.join(output_dir, f"temp_tbm_{date_str}.xlsx").replace("/", "\\")
        
        excel = None
        try:
            # 0. 날짜 및 UI 동기화 (메인 탭 기준)
            main_date = self.ent_date.get().strip()
            
            if hasattr(self, 'tbm_manager'):
                self.tbm_manager.ent_date.delete(0, tk.END)
                self.tbm_manager.ent_date.insert(0, main_date)
                self.tbm_manager.update_work_and_hazards()
                
            if hasattr(self, 'risk_manager'):
                try:
                    import datetime
                    dt = datetime.datetime.strptime(main_date, "%Y-%m-%d")
                    risk_date = dt.strftime("%Y년 %m월 %d일")
                    self.risk_manager.ent_write_date.delete(0, tk.END)
                    self.risk_manager.ent_write_date.insert(0, risk_date)
                except:
                    pass

            # 1. 작업승인계획서 생성
            self.generate_files(silent_path=temp_approval)
            
            # 2. TBM 생성
            if hasattr(self, 'tbm_manager'):
                self.tbm_manager.export_excel(silent_path=temp_tbm)
                
            # 3. 위험성평가표 생성
            temp_risks = []
            if hasattr(self, 'risk_manager'):
                temp_risks = self.risk_manager.generate_files(silent_dir=output_dir, date_str=date_str) or []
                
            # 합치기
            import win32com.client as win32
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # 새 통합 워크북 생성
            wb_master = excel.Workbooks.Add()
            default_sheet = wb_master.Sheets(1)
            
            # 작업승인계획서 복사 (Before=default_sheet)
            if os.path.exists(temp_approval):
                wb1 = excel.Workbooks.Open(temp_approval)
                wb1.Sheets(1).Copy(wb_master.Sheets(1))
                wb1.Close(False)
                try: os.remove(temp_approval)
                except: pass
                
            # TBM 복사 (After=첫번째 시트)
            if os.path.exists(temp_tbm):
                wb2 = excel.Workbooks.Open(temp_tbm)
                wb2.Sheets(1).Copy(None, wb_master.Sheets(1))
                wb2.Close(False)
                try: os.remove(temp_tbm)
                except: pass
                
            # 위험성평가표 복사 (After=마지막 시트)
            for temp_risk in temp_risks:
                if os.path.exists(temp_risk):
                    last_sheet = wb_master.Sheets(wb_master.Sheets.Count)
                    wb3 = excel.Workbooks.Open(temp_risk)
                    wb3.Sheets(1).Copy(None, last_sheet)
                    wb3.Close(False)
                    try: os.remove(temp_risk)
                    except: pass
                    
            # 기본 Sheet1 삭제
            default_sheet.Delete()
            
            wb_master.SaveAs(final_path)
            wb_master.Close(False)
            excel.Quit()
            
            messagebox.showinfo("통합 완료", f"일일 안전서류 통합 엑셀이 성공적으로 생성되었습니다!\n\n저장 위치:\n{final_path}")
            self.lbl_status.config(text="통합 완료!", foreground="green")
            os.startfile(final_path)
            
        except Exception as e:
            messagebox.showerror("오류", f"통합 엑셀 생성 중 오류가 발생했습니다:\n{e}")
            self.lbl_status.config(text="통합 오류", foreground="red")
            try:
                if excel: excel.Quit()
            except: pass
        finally:
            self.btn_generate.config(state='normal')
            self.btn_unified.config(state='normal')

    def create_excel(self, output_path, params):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "작업승인계획서"

        # Fonts & Styles
        title_font = Font(name='맑은 고딕', bold=True, size=16)
        subtitle_font = Font(name='맑은 고딕', bold=True, size=12)
        bold_font = Font(name='맑은 고딕', bold=True, size=10)
        normal_font = Font(name='맑은 고딕', size=10)
        small_font = Font(name='맑은 고딕', size=9, bold=True, color='0000FF')
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        def set_border(ws, min_col, min_row, max_col, max_row):
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.border = thin_border
                    
        def write_cell(c, r, val, font=normal_font, align=center_align):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = font
            cell.alignment = align
            return cell

        # Column widths
        cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        widths = [5, 10, 10, 12, 10, 12, 10, 10, 12, 12, 12]
        for c, w in zip(cols, widths):
            ws.column_dimensions[c].width = w

        # Header Section
        ws.merge_cells('A1:E1')
        ws['A1'] = "[서식 3] 작업승인계획서"
        ws['A1'].font = bold_font
        ws['A1'].alignment = left_align
        
        ws.merge_cells('C2:H3')
        write_cell(3, 2, "작업승인계획서", title_font)
        
        w_type = params.get('work_type', '위험')
        r_mark = "■" if w_type == "위험" else "□"
        n_mark = "■" if w_type == "일반" else "□"
        type_str = f"( {r_mark} 위험   {n_mark} 일반 )"
        ws.merge_cells('C4:H4')
        write_cell(3, 4, type_str, subtitle_font)

        # Signature Block
        ws.merge_cells('F5:G6')
        write_cell(6, 5, "수급업체\n현장대리인", bold_font)
        ws.merge_cells('H5:I6')
        write_cell(8, 5, "시공감독\n(KOGAS)", bold_font)
        ws.merge_cells('J5:K6')
        write_cell(10, 5, "소장\n(KOGAS)", bold_font)
        
        ws.merge_cells('F7:G9')
        ws.merge_cells('H7:I9')
        ws.merge_cells('J7:K9')
        
        write_cell(6, 7, "")
        write_cell(8, 7, "전자결재\n(안전달인)")
        write_cell(10, 7, "전자결재\n(안전달인)")
        
        ws.merge_cells('F10:G11')
        write_cell(6, 10, "KOGAS 안전감독\n(위험작업시)", bold_font)
        ws.merge_cells('H10:K11')
        write_cell(8, 10, "전자결재 (안전달인)")
        
        set_border(ws, 6, 5, 11, 11)

        # Risk Types Block
        ws.merge_cells('A5:C5')
        write_cell(1, 5, "※ 위험작업종류", small_font, left_align)
        
        risks = params.get('risks', {})
        def g_risk(key): return "■" if risks.get(key, False) else "□"
        
        risk_map = [
            [f"{g_risk('화기작업')}화기작업", f"{g_risk('밀폐공간 작업')}밀폐공간 작업", f"{g_risk('고소작업')}고소작업"],
            [f"{g_risk('전기통신 작업')}전기통신 작업", f"{g_risk('화학물질 취급작업')}화학물질 취급작업", f"{g_risk('굴착작업')}굴착작업"],
            [f"{g_risk('중량물 취급작업')}중량물 취급작업", f"{g_risk('지하공간 가스방출')}지하공간 가스방출", f"{g_risk('지붕위·철골작업')}지붕위·철골작업"],
            [f"{g_risk('기타 위험작업')}기타 위험작업", "", ""]
        ]
        
        for r_idx, r_list in enumerate(risk_map):
            row = 6 + r_idx
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            write_cell(1, row, r_list[0], small_font, left_align)
            
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            write_cell(3, row, r_list[1], small_font, left_align)
            
            write_cell(5, row, r_list[2], small_font, left_align)
            
        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 25
        ws.row_dimensions[4].height = 20
        for i in range(5, 12):
            ws.row_dimensions[i].height = 20
            
        # Form Body
        r = 13
        # 작업명, 기간
        ws.merge_cells(f'A{r}:B{r}')
        write_cell(1, r, "작업명", bold_font)
        ws.merge_cells(f'C{r}:F{r}')
        write_cell(3, r, params.get('work_name', ''), align=left_align)
        
        ws.merge_cells(f'G{r}:H{r}')
        write_cell(7, r, "작업기간", bold_font)
        ws.merge_cells(f'I{r}:K{r}')
        write_cell(9, r, params.get('work_period', '          ~          '))
        ws.row_dimensions[r].height = 30
        set_border(ws, 1, r, 11, r)
        r += 1

        # 작업수행부서
        ws.merge_cells(f'A{r}:B{r+1}')
        write_cell(1, r, "작업수행부서", bold_font)
        
        ws.merge_cells(f'C{r}:D{r}')
        write_cell(3, r, "원도급 업체명")
        ws.merge_cells(f'E{r}:F{r}')
        write_cell(5, r, params.get('main_company', ''), align=left_align)
        ws.merge_cells(f'G{r}:H{r}')
        write_cell(7, r, "작업책임자")
        ws.merge_cells(f'I{r}:K{r}')
        write_cell(9, r, params.get('main_manager', ''))
        
        ws.merge_cells(f'C{r+1}:D{r+1}')
        write_cell(3, r+1, "하도급 업체명")
        ws.merge_cells(f'E{r+1}:F{r+1}')
        write_cell(5, r+1, params.get('sub_company', ''), align=left_align)
        ws.merge_cells(f'G{r+1}:H{r+1}')
        write_cell(7, r+1, "작업책임자")
        ws.merge_cells(f'I{r+1}:K{r+1}')
        write_cell(9, r+1, params.get('sub_manager', ''))
        
        ws.row_dimensions[r].height = 30
        ws.row_dimensions[r+1].height = 30
        set_border(ws, 1, r, 11, r+1)
        r += 2

        # 투입현황
        ws.merge_cells(f'A{r}:B{r}')
        write_cell(1, r, "작업인원", bold_font)
        ws.merge_cells(f'C{r}:K{r}')
        write_cell(3, r, params.get('personnel', '총원    명 (직종별 인원 세부적으로 작성)'), align=left_align)
        ws.row_dimensions[r].height = 40
        set_border(ws, 1, r, 11, r)
        r += 1
        
        ws.merge_cells(f'A{r}:B{r}')
        write_cell(1, r, "사용장비", bold_font)
        ws.merge_cells(f'C{r}:K{r}')
        write_cell(3, r, params.get('equip', '(종류, 수량 기재)'), align=left_align)
        ws.row_dimensions[r].height = 30
        set_border(ws, 1, r, 11, r)
        r += 1
        
        ws.merge_cells(f'A{r}:B{r}')
        write_cell(1, r, "개인보호장비", bold_font)
        ws.merge_cells(f'C{r}:K{r}')
        write_cell(3, r, params.get('ppe', '(종류, 수량 기재)'), align=left_align)
        ws.row_dimensions[r].height = 30
        set_border(ws, 1, r, 11, r)
        r += 1
        
        ws.merge_cells(f'A{r}:B{r}')
        write_cell(1, r, "기타 안전장비", bold_font)
        ws.merge_cells(f'C{r}:K{r}')
        write_cell(3, r, params.get('safety_equip', '(종류, 수량 기재)'), align=left_align)
        ws.row_dimensions[r].height = 30
        set_border(ws, 1, r, 11, r)
        r += 1

        # 작업내용
        ws.merge_cells(f'A{r}:K{r}')
        write_cell(1, r, "작업내용 (일일 수행 모든 작업사항을 세부적으로 작성)", bold_font, left_align)
        ws.row_dimensions[r].height = 20
        set_border(ws, 1, r, 11, r)
        r += 1
        
        ws.merge_cells(f'A{r}:K{r}')
        write_cell(1, r, params.get('work_content', ''), align=Alignment(horizontal="left", vertical="top", wrap_text=True))
        ws.row_dimensions[r].height = 250
        set_border(ws, 1, r, 11, r)
        r += 1

        # 주요 위험요인
        ws.merge_cells(f'A{r}:B{r}')
        write_cell(1, r, "주요 위험요인", bold_font)
        ws.merge_cells(f'C{r}:K{r}')
        write_cell(3, r, params.get('risk_factors', ''), align=Alignment(horizontal="left", vertical="top", wrap_text=True))
        ws.row_dimensions[r].height = 150
        set_border(ws, 1, r, 11, r)
        r += 1

        # 안전부서 검토의견
        ws.merge_cells(f'A{r}:B{r}')
        write_cell(1, r, "안전부서\n검토의견", bold_font)
        ws.merge_cells(f'C{r}:K{r}')
        write_cell(3, r, "(위험작업 시 안전부 검토)", Font(name='맑은 고딕', size=10, italic=True, color='808080'))
        ws.row_dimensions[r].height = 100
        set_border(ws, 1, r, 11, r)
        r += 1

        # 첨부
        ws.merge_cells(f'A{r}:B{r}')
        write_cell(1, r, "첨부", bold_font)
        ws.merge_cells(f'C{r}:K{r}')
        write_cell(3, r, "위험성평가 결과, 도면, 시방서 등 작업관련 자료 첨부", align=left_align)
        ws.row_dimensions[r].height = 30
        set_border(ws, 1, r, 11, r)
        r += 1

        # 현장점검 체크리스트
        ws.merge_cells(f'A{r}:B{r}')
        write_cell(1, r, "현장점검\n체크리스트", bold_font)
        ws.merge_cells(f'C{r}:K{r}')
        write_cell(3, r, params.get('checklist', '전기작업, 굴착작업, 화기작업, 밀폐공간 작업시 작업책임자가 현장에서 체크리스트 점검사항 직접 확인 후 기입'), align=left_align)
        ws.row_dimensions[r].height = 80
        set_border(ws, 1, r, 11, r)

        # Print settings
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_options.horizontalCentered = True

        wb.save(output_path)

if __name__ == "__main__":
    root = tk.Tk()
    app = WorkApprovalApp(root)
    root.mainloop()
