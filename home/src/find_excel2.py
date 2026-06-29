import os
import openpyxl

desktop = r'C:\Users\-\OneDrive\바탕 화면'
files = [f for f in os.listdir(desktop) if '산출내역서' in f and f.endswith('.xlsx') and not f.startswith('~$')]
print("Found files:", files)

if files:
    filepath = os.path.join(desktop, files[0])
    print(f"Reading {filepath}...")
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        print("Sheets:", wb.sheetnames)
        
        for sheet_name in wb.sheetnames:
            if 'UT' in sheet_name or 'PT' in sheet_name:
                print(f"\n--- {sheet_name} ---")
                ws = wb[sheet_name]
                for r in range(1, min(100, ws.max_row + 1)):
                    row_data = []
                    for c in range(1, min(20, ws.max_column + 1)):
                        val = ws.cell(row=r, column=c).value
                        if val is not None:
                            # Clean up newlines for printing
                            if isinstance(val, str):
                                val = val.replace('\n', ' ')
                            row_data.append(str(val))
                        else:
                            row_data.append("")
                    if any(row_data):
                        print(f"R{r}:", "\t".join(row_data))
    except Exception as e:
        print("Error:", e)
