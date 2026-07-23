import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import os
from datetime import datetime

class InfoSurveyApp:
    def __init__(self, root):
        self.root = root
        self.root.title("[붙임4] 안전보건정보 조사 자동 생성기")
        self.root.geometry("450x250")
        
        style = ttk.Style()
        style.theme_use('clam')
        
        main_frame = ttk.Frame(self.root, padding=20)
        main_frame.pack(fill='both', expand=True)
        
        ttk.Label(main_frame, text="안전보건정보 조사(붙임4) 생성", font=('Malgun Gothic', 14, 'bold')).pack(pady=(0, 20))
        
        form_frame = ttk.LabelFrame(main_frame, text="입력 정보", padding=15)
        form_frame.pack(fill='x', pady=5)
        
        ttk.Label(form_frame, text="작성일:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.ent_date = ttk.Entry(form_frame, width=25)
        today = datetime.now().strftime("%Y년 %m월 %d일")
        self.ent_date.insert(0, today)
        self.ent_date.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(form_frame, text="작성자:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.ent_author = ttk.Entry(form_frame, width=25)
        self.ent_author.insert(0, "주진철")
        self.ent_author.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=15)
        
        self.btn_generate = ttk.Button(btn_frame, text="엑셀 파일 생성", command=self.generate_excel)
        self.btn_generate.pack(fill='x', ipady=5)
        
        self.lbl_status = ttk.Label(main_frame, text="대기 중...", foreground="gray")
        self.lbl_status.pack()

    def generate_excel(self):
        self.btn_generate.config(state='disabled')
        self.lbl_status.config(text="생성 중...", foreground="blue")
        self.root.update()
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "안전보건정보조사"
            
            # --- 페이지 설정 ---
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.6
            ws.page_margins.bottom = 0.6
            ws.print_options.horizontalCentered = True
            
            # --- 열 너비 설정 (A~H) ---
            widths = [15, 18, 7, 15, 10, 10, 18, 32]
            for i, width in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
                
            # --- 스타일 ---
            normal_font = Font(name='맑은 고딕', size=10)
            blue_font = Font(name='맑은 고딕', size=11, bold=True, color="0000FF")
            title_font = Font(name='맑은 고딕', size=22, bold=True, underline="single")
            fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            thin = Side(style='thin')
            thick = Side(style='medium')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            def set_border(min_col, min_row, max_col, max_row, style=border):
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        ws.cell(row=r, column=c).border = style

            # --- 상단 타이틀 ---
            ws.merge_cells('A1:H1')
            ws['A1'] = "[붙임4] 안전보건정보조사(활용 정보 목록)"
            ws['A1'].font = blue_font
            ws['A1'].alignment = left_align
            ws.row_dimensions[1].height = 20
            
            ws.merge_cells('A2:H2')
            ws['A2'] = "안전보건정보 조사"
            ws['A2'].font = title_font
            ws['A2'].alignment = center_align
            ws.row_dimensions[2].height = 40
            ws.row_dimensions[3].height = 10
            
            # --- 기본 틀 그리기 ---
            set_border(1, 4, 8, 21)
            
            # Row 4, 5
            ws['A4'] = "작업(공정)"
            ws.merge_cells('B4:D4')
            ws['B4'] = "비파괴검사 기술용역"
            ws.merge_cells('E4:F5')
            ws['E4'] = "안전보건정보"
            ws['E4'].font = Font(name='맑은 고딕', size=18, bold=True)
            ws['G4'] = "작성일"
            ws['H4'] = self.ent_date.get()
            
            ws['A5'] = "근로자수"
            ws.merge_cells('B5:D5')
            ws['G5'] = "작성자"
            ws['H5'] = self.ent_author.get()
            
            for r in [4, 5]:
                ws.row_dimensions[r].height = 25
                for c in range(1, 9):
                    ws.cell(row=r, column=c).alignment = center_align
                    if ws.cell(row=r, column=c).font.size != 18:
                        ws.cell(row=r, column=c).font = normal_font
                        ws.cell(row=r, column=c).fill = fill_gray if c not in [2, 3, 4, 8] else PatternFill(fill_type=None)
            ws['B4'].fill = PatternFill(fill_type=None) # 예외 처리
            
            # Row 6, 7 헤더
            ws.merge_cells('A6:A7')
            ws['A6'] = "공정(작업)순서"
            ws.merge_cells('B6:C6')
            ws['B6'] = "기계·기구 및 설비"
            ws.merge_cells('D6:F6')
            ws['D6'] = "유해(화학)물질"
            ws.merge_cells('G6:H7')
            ws['G6'] = "그 밖의 유해위험정보"
            
            ws['B7'] = "기계기구 및 설비명"
            ws['C7'] = "수량"
            ws['D7'] = "화학물질명"
            ws['E7'] = "취급량/일"
            ws['F7'] = "취급시간"
            
            for r in [6, 7]:
                ws.row_dimensions[r].height = 20
                for c in range(1, 9):
                    ws.cell(row=r, column=c).alignment = center_align
                    ws.cell(row=r, column=c).font = normal_font
                    ws.cell(row=r, column=c).fill = fill_gray
            
            # --- 좌측 8단계 공정 (Row 8 ~ 21) ---
            processes = [
                ("작업준비", "서베이미터\nTLD\n알람도시메터", "1", "", "", "0.2hr", 8, 9),
                ("폭염 및 혹서기", "물, 식염수, 포도당 및\n그늘/바람, 휴식", "", "", "", "", 10, 10),
                ("혹한기", "방한장구, 따뜻한물, 휴게실", "", "", "", "", 11, 11),
                ("고령근로자", "휴식 및 순환근무,\n근골격계 질환 관리\n(작업전.후 스트레칭)", "", "방사선조사기", "20kg/d", "0.2hr", 12, 13),
                ("방사선투과검사(RT)", "안전표지판, 안전경계용 줄,\n경고등", "1", "방사선", "100매/d", "5hr", 14, 15),
                ("현상작업", "앞치마,고무장갑,보안경", "1", "디하이드록시벤젠\n글루타르알데히드", "100매/d", "1hr", 16, 17),
                ("침투탐상검사(PT)", "방독마스크,\n보안경", "1", "이소프로필 알코올\n헵탄\n활석\n톨루엔", "450ml/d", "1hr", 18, 19),
                ("현장 정리 및\n폐기물 보관", "서베이미터\n알람도시메터\n저장탱크", "1", "", "", "0.2hr", 20, 21)
            ]
            
            for p_name, m_name, qty, chem, c_qty, c_time, start_row, end_row in processes:
                ws.merge_cells(f'A{start_row}:A{end_row}')
                ws[f'A{start_row}'] = p_name
                ws.merge_cells(f'B{start_row}:B{end_row}')
                ws[f'B{start_row}'] = m_name
                ws.merge_cells(f'C{start_row}:C{end_row}')
                ws[f'C{start_row}'] = qty
                ws.merge_cells(f'D{start_row}:D{end_row}')
                ws[f'D{start_row}'] = chem
                ws.merge_cells(f'E{start_row}:E{end_row}')
                ws[f'E{start_row}'] = c_qty
                ws.merge_cells(f'F{start_row}:F{end_row}')
                ws[f'F{start_row}'] = c_time
            
            for r in range(8, 22):
                for c in range(1, 7):
                    ws.cell(row=r, column=c).alignment = center_align
                    ws.cell(row=r, column=c).font = normal_font
            
            # --- 우측 그밖의 정보 (G, H열 매핑) ---
            right_items_with_header = [
                ("활용정보", "문서 제목"),
                ("작업표준\n작업절차 등", "용역시방서 참조\n비파괴검사 표준 작업절차등 참고"),
                ("MSDS/\n기계·설비 사양서", "RT약품 MSDS 참조 / PT약품 MSDS 참조\nCERTIFICATE of Sealed Source 참조"),
                ("기계·기구, 설비\n등, 주변 환경정보", "관리소 도면, 용역시방서 참조"),
                ("재해사례,\n재해통계 정보", "비파괴검사 중 낙하사고\n작업이동중 차량사고 등 참조")
            ]
            for i, (col1, col2) in enumerate(right_items_with_header):
                r = 8 + i
                ws[f'G{r}'] = col1
                ws[f'H{r}'] = col2
                if r == 8:
                    ws.cell(row=r, column=7).alignment = center_align
                    ws.cell(row=r, column=8).alignment = center_align
                else:
                    ws.cell(row=r, column=7).alignment = center_align
                    ws.cell(row=r, column=8).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.cell(row=r, column=7).font = normal_font if r != 8 else Font(name='맑은 고딕', size=10, bold=True, underline="single")
                ws.cell(row=r, column=8).font = normal_font if r != 8 else Font(name='맑은 고딕', size=10, bold=True)
            
            ws.merge_cells('G13:H13')
            ws['G13'] = "○ 작업환경측정 측정유무(측정□, 미측정□, 해당무■)"
            ws.merge_cells('G14:H14')
            ws['G14'] = "○ 도급(일부, 전부 또는 혼재작업) (유□, 무■)"
            ws.merge_cells('G15:H15')
            ws['G15'] = "○ 근로자 건강진단 유무 (유■, 무□)"
            ws.merge_cells('G16:H16')
            ws['G16'] = "○ 안전작업허가증 필요작업 유무(유■, 무□)"
            
            ws.merge_cells('G17:H17')
            ws['G17'] = "○ 중량물 인력취급시 단위중량(20kg) 및 취급형태\n    (들기■, 밀기□, 끌기□)"
            
            ws.merge_cells('G18:H18')
            ws['G18'] = "○ 근로자 구성 및 경력특성"
            
            ws['G19'] = "여성 근로자           □"
            ws['H19'] = "1년 미만 미숙련자   □"
            ws['G20'] = "고령 근로자           □"
            ws['H20'] = "비정규직 근로자      □"
            ws['G21'] = "외국인 근로자        □"
            ws['H21'] = "장애 근로자           □"
            
            # Formatting checkboxes
            for r in range(13, 22):
                if r in [19, 20, 21]:
                    ws.cell(row=r, column=7).alignment = Alignment(horizontal='left', vertical='center')
                    ws.cell(row=r, column=8).alignment = Alignment(horizontal='left', vertical='center')
                    ws.cell(row=r, column=7).border = border
                    ws.cell(row=r, column=8).border = border
                else:
                    ws.cell(row=r, column=7).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.cell(row=r, column=7).font = normal_font
                ws.cell(row=r, column=8).font = normal_font
            
            # "그 밖에 위험성평가" 행은 22로 밀림 (공정은 21에서 끝나지만 그밖에는 22에 위치)
            # 사실 22행까지 표에 포함시킵니다.
            ws.merge_cells('A22:F22')
            ws.merge_cells('G22:H22')
            ws['G22'] = "○ 그 밖에 위험성평가에 참고가 되는 자료 등"
            ws['G22'].font = normal_font
            ws['G22'].alignment = Alignment(horizontal='left', vertical='center')
            
            # Row heights
            ws.row_dimensions[8].height = 20
            ws.row_dimensions[9].height = 30
            ws.row_dimensions[10].height = 30
            ws.row_dimensions[11].height = 25
            ws.row_dimensions[12].height = 25
            
            for r in range(13, 17): ws.row_dimensions[r].height = 20
            ws.row_dimensions[17].height = 30 # 중량물
            ws.row_dimensions[18].height = 20 # 구성
            ws.row_dimensions[19].height = 18
            ws.row_dimensions[20].height = 18
            ws.row_dimensions[21].height = 18
            ws.row_dimensions[22].height = 20
            
            # Outer thick borders
            for c in range(1, 9): ws.cell(row=4, column=c).border = Border(top=thick, left=thin, right=thin, bottom=thin)
            ws.cell(row=4, column=1).border = Border(top=thick, left=thick, right=thin, bottom=thin)
            ws.cell(row=4, column=8).border = Border(top=thick, left=thin, right=thick, bottom=thin)
            
            for r in range(5, 22):
                ws.cell(row=r, column=1).border = Border(top=thin, left=thick, right=thin, bottom=thin)
                ws.cell(row=r, column=8).border = Border(top=thin, left=thin, right=thick, bottom=thin)
            
            for c in range(1, 9): ws.cell(row=22, column=c).border = Border(top=thin, left=thin, right=thin, bottom=thick)
            ws.cell(row=22, column=1).border = Border(top=thin, left=thick, right=thin, bottom=thick)
            ws.cell(row=22, column=8).border = Border(top=thin, left=thin, right=thick, bottom=thick)
            
            for r in [19, 20, 21]:
                ws.cell(row=r, column=7).border = Border(left=thick, right=thin, top=thin, bottom=thin)
                ws.cell(row=r, column=8).border = Border(left=thin, right=thick, top=thin, bottom=thin)
            ws.cell(row=19, column=7).border = Border(left=thick, right=thin, top=thick, bottom=thin)
            ws.cell(row=19, column=8).border = Border(left=thin, right=thick, top=thick, bottom=thin)
            ws.cell(row=21, column=7).border = Border(left=thick, right=thin, top=thin, bottom=thick)
            ws.cell(row=21, column=8).border = Border(left=thin, right=thick, top=thin, bottom=thick)

            # --- Footer ---
            ws.merge_cells('A23:H23')
            ws['A23'] = "※ 유해화학물질 : 법 제39조 제1항에 따라 고용노동부령으로 정하는 분류기준(시행규칙 별표11의2)에 해당하는 화학물질 및 화학물질을 함유한 제제"
            ws['A23'].font = Font(name='맑은 고딕', size=9)
            ws['A23'].alignment = Alignment(horizontal='left', vertical='top')
            ws.row_dimensions[23].height = 20
            
            ws.merge_cells('A25:H25')
            ws['A25'] = "15"
            ws['A25'].font = normal_font
            ws['A25'].alignment = Alignment(horizontal='center', vertical='bottom')
            
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=f"안전보건정보조사_{datetime.now().strftime('%Y%m%d')}.xlsx",
                title="엑셀 파일 저장",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if save_path:
                wb.save(save_path)
                messagebox.showinfo("완료", "안전보건정보조사 엑셀 파일이 성공적으로 생성되었습니다!")
                self.lbl_status.config(text="생성 완료!", foreground="green")
            else:
                self.lbl_status.config(text="저장 취소됨", foreground="gray")
                
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 생성 중 오류 발생:\n{e}")
            self.lbl_status.config(text="오류 발생", foreground="red")
        finally:
            self.btn_generate.config(state='normal')

if __name__ == "__main__":
    root = tk.Tk()
    app = InfoSurveyApp(root)
    root.mainloop()
