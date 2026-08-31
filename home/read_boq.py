import openpyxl
import os

file_path = r'C:\Users\-\OneDrive\바탕 화면\산출내역서(가산~가평 천연가스 공급시설 건설공사 비파괴검사 기술용역)_감독경유.xlsx'

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
else:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(f"Sheet names: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames[:3]: # Limit to first 3 sheets to avoid too much output
            ws = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")
            for r in range(1, 20):
                row_vals = []
                for c in range(1, 12):
                    val = ws.cell(row=r, column=c).value
                    row_vals.append(str(val) if val is not None else "")
                if any(row_vals): # Only print non-empty rows
                    print(f"Row {r}: " + " | ".join(row_vals))
    except Exception as e:
        print(f"Error reading excel file: {e}")
