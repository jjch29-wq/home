import openpyxl

excel_path = r"C:\Users\-\OneDrive\바탕 화면\템플릿 가산가평월간용역진도보고서.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

def analyze_sheet(ws_name):
    ws = wb[ws_name]
    totals = {"RT_B": 0, "RT_A": 0, "RT_A2": 0, "UT": 0, "PT": 0}
    current_section = None
    
    for r in range(4, 150):
        val1 = ws.cell(row=r, column=1).value
        if isinstance(val1, str):
            if "방사선투과검사" in val1: current_section = "RT"
            elif "초음파탐상" in val1 or "UT" in val1.upper(): current_section = "UT"
            elif "액체침투탐상" in val1 or "PT" in val1.upper(): current_section = "PT"
        
        if current_section == "RT":
            # RT uses Col 8 for film type and Col 7 for total films
            film_type = ws.cell(row=r, column=8).value
            film_qty = ws.cell(row=r, column=7).value
            if isinstance(film_qty, (int, float)) and isinstance(film_type, str):
                if "B-TYPE" in film_type: totals["RT_B"] += film_qty
                elif "A-TYPE" in film_type: totals["RT_A"] += film_qty
                elif "A/2-TYPE" in film_type: totals["RT_A2"] += film_qty
        
        elif current_section in ["UT", "PT"]:
            # UT/PT uses Col 7 for total length or Col 5 for 실검사길이 and Col 7 for 검사보정길이
            # Let's sum Col 5 (실검사길이)
            qty = ws.cell(row=r, column=5).value
            if isinstance(qty, (int, float)):
                totals[current_section] += qty
                
    return totals

print("--- 주배관 ---")
print(analyze_sheet("나. 검사물량 세부내역 (주배관)"))

print("--- 관리소 ---")
print(analyze_sheet("나. 검사물량 세부내역 (관리소)"))
