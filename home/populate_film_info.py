import openpyxl
from openpyxl.cell.cell import MergedCell

def get_dia_val(dia_str):
    if not dia_str: return None
    dia_str = str(dia_str).replace('"', '').replace("'", '').replace('(wed)', '').strip()
    if not dia_str: return None
    try:
        if '/' in dia_str:
            num, den = dia_str.split('/')
            return float(num) / float(den)
        return float(dia_str)
    except:
        return None

def get_shots(dia):
    if dia >= 30: return 7
    elif dia >= 14: return 5
    elif dia >= 10: return 4
    elif dia >= 6: return 3
    elif dia == 4: return 4
    elif dia == 2.5: return 3
    else: return 4

def get_film_type(dia):
    if dia >= 14: return 'B-TYPE'
    elif dia > 4: return 'A-TYPE'
    else: return 'A/2-TYPE'

def write_val(ws, r, c, val):
    cell = ws.cell(row=r, column=c)
    if isinstance(cell, MergedCell):
        # Find the top-left cell of the merged range this cell belongs to
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left_cell.value = val
                return
    else:
        cell.value = val

excel_path = r"C:\Users\-\OneDrive\바탕 화면\템플릿 가산가평월간용역진도보고서.xlsx"
wb = openpyxl.load_workbook(excel_path)

sheets_to_update = ["나. 검사물량 세부내역 (주배관)", "나. 검사물량 세부내역 (관리소)"]

for sheet_name in sheets_to_update:
    if sheet_name not in wb.sheetnames: continue
    ws = wb[sheet_name]
    for r in range(6, 45): # Iterate through possible rows
        dia_cell = ws.cell(row=r, column=2).value
        dia_val = get_dia_val(dia_cell)
        
        if dia_val is not None:
            shots = get_shots(dia_val)
            film_type = get_film_type(dia_val)
            
            write_val(ws, r, 3, shots)
            write_val(ws, r, 8, film_type)

wb.save(excel_path)
print("Successfully populated shots and film types.")
