import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

output_path = r"C:\Users\-\OneDrive\바탕 화면\4.4.1_위험성평가표(RT)_최종_수정본.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "합동 위험성평가표"

# Styles
bold_font = Font(bold=True)
title_font = Font(bold=True, size=18)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

def set_border(ws, min_col, min_row, max_col, max_row):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border

# Use 20 columns to decouple header widths from data widths.
# Base width = 4
for i in range(1, 25):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 4

# --- Header Section (Rows 1 to 9) ---
ws.merge_cells('A1:T1')
ws['A1'] = "합동 위험성평가표 [ ■최초, □정기, □수시 ]"
ws['A1'].font = title_font
ws['A1'].alignment = center_align
ws.row_dimensions[1].height = 40

ws.merge_cells('A2:B3')
ws['A2'] = "작 업 명"
ws.merge_cells('C2:E3')
ws['C2'] = "가산~가평 천연가스 공급설비\n건설공사 비파괴검사 기술용역"
ws.merge_cells('F2:G2')
ws['F2'] = "평가기간"
ws.merge_cells('H2:M2')
ws['H2'] = "2026. 07. 09"
ws.merge_cells('N2:P2')
ws['N2'] = "재평가일"
ws.merge_cells('Q2:T2')

ws.merge_cells('A4:B5')
ws['A4'] = "회 사 명"
ws.merge_cells('C4:E5')
ws['C4'] = "서울검사(주)"

ws.merge_cells('A6:B7')
ws['A6'] = "작업공종"
ws.merge_cells('C6:E7')
ws['C6'] = "방사선투과검사"

ws.merge_cells('A8:B9')
ws['A8'] = "작업기간"
ws.merge_cells('C8:E9')
ws['C8'] = "2026.06.22 ~ 2028.10.31"

ws.merge_cells('F3:F9')
ws['F3'] = "작업전"

ws.merge_cells('G3:G5')
ws['G3'] = "수급인"

# Equal widths for signatures: H~I (2), J~K (2), L~M (2)
ws.merge_cells('H3:I3')
ws['H3'] = "근 로 자"
ws.merge_cells('J3:K3')
ws['J3'] = "작 성 자"
ws.merge_cells('L3:M3')
ws['L3'] = "승 인 자"

ws.merge_cells('H4:I5')
ws['H4'] = "유상훈 (서명)"
ws.merge_cells('J4:K5')
ws['J4'] = "주진철 (서명)"
ws.merge_cells('L4:M5')
ws['L4'] = "강신태 (서명)"

ws.merge_cells('G6:G9')
ws['G6'] = "도급인"
ws.merge_cells('H6:I6')
ws['H6'] = "검토자(감독)"
ws.merge_cells('J6:K6')
ws['J6'] = "검토자(안전)"
ws.merge_cells('L6:M6')
ws['L6'] = "승 인 자"
ws.merge_cells('H7:I9')
ws['H7'] = "(서명)"
ws.merge_cells('J7:K9')
ws['J7'] = "(서명)"
ws.merge_cells('L7:M9')
ws['L7'] = "(서명)"

# 재평가 signatures
ws.merge_cells('N3:N9')
ws['N3'] = "재평가"

ws.merge_cells('O3:O5')
ws['O3'] = "수급인"

ws.merge_cells('P3:Q3')
ws['P3'] = "근 로 자"
ws.merge_cells('R3:S3')
ws['R3'] = "작 성 자"
ws.merge_cells('T3:T3')
ws['T3'] = "승 인 자"
# Wait, let's adjust the grid. 
# T is just 1 column. Let's make P:Q(2), R:S(2), T:U(2). So we need U.
for i in range(1, 23):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 4
ws.merge_cells('A1:U1') # expand title
ws.merge_cells('Q2:U2') # expand 재평가일 blank

ws.merge_cells('T3:U3')
ws['T3'] = "승 인 자"

ws.merge_cells('P4:Q5')
ws['P4'] = "(서명)"
ws.merge_cells('R4:S5')
ws['R4'] = "(서명)"
ws.merge_cells('T4:U5')
ws['T4'] = "(서명)"

ws.merge_cells('O6:O9')
ws['O6'] = "도급인"
ws.merge_cells('P6:Q6')
ws['P6'] = "검토자(감독)"
ws.merge_cells('R6:S6')
ws['R6'] = "검토자(안전)"
ws.merge_cells('T6:U6')
ws['T6'] = "승 인 자"
ws.merge_cells('P7:Q9')
ws['P7'] = "(서명)"
ws.merge_cells('R7:S9')
ws['R7'] = "(서명)"
ws.merge_cells('T7:U9')
ws['T7'] = "(서명)"

for r in range(2, 10):
    for c in range(1, 22):
        cell = ws.cell(row=r, column=c)
        cell.alignment = center_align
        cell.font = Font(size=10)

set_border(ws, 1, 2, 21, 9)

# --- Table Sub-Header (Rows 10-12) ---
ws.merge_cells('A10:M10')
ws['A10'] = "작업 전 평가"
ws.merge_cells('N10:U10')
ws['N10'] = "재평가"

ws.merge_cells('A11:A12')
ws['A11'] = "세부작업"
ws.merge_cells('B11:E11')
ws['B11'] = "유해 위험 요인 파악"
ws.merge_cells('F11:G12')
ws['F11'] = "현재의 안전 보건 조치"
ws.merge_cells('H11:J11')
ws['H11'] = "위험성 결정"
ws.merge_cells('K11:O12')
ws['K11'] = "위험성 감소 대책"
ws.merge_cells('P11:P12')
ws['P11'] = "개선\n예정일"
ws.merge_cells('Q11:Q12')
ws['Q11'] = "담당자\n(담당부서)"
ws.merge_cells('R11:R12')
ws['R11'] = "완료일"
ws.merge_cells('S11:U11')
ws['S11'] = "개선 후 위험성 (재평가)"

ws.merge_cells('B12:B12')
ws['B12'] = "위험\n분류"
ws.merge_cells('C12:E12')
ws['C12'] = "위험 발생 상황 및 결과"
ws['H12'] = "가능성\n(빈도)"
ws['I12'] = "중대성\n(강도)"
ws['J12'] = "위험성\n(등급)"
ws['S12'] = "가능성\n(빈도)"
ws['T12'] = "중대성\n(강도)"
ws['U12'] = "위험성\n(등급)"

for r in range(10, 13):
    for c in range(1, 22):
        cell = ws.cell(row=r, column=c)
        cell.alignment = center_align
        cell.font = bold_font
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

set_border(ws, 1, 10, 21, 12)

# Specific Column Width Adjustments to make it look right
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 6
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 7 # 근로자
ws.column_dimensions['I'].width = 7 # 근로자
ws.column_dimensions['J'].width = 7 # 작성자
ws.column_dimensions['K'].width = 7 # 작성자
ws.column_dimensions['L'].width = 7 # 승인자
ws.column_dimensions['M'].width = 7 # 승인자

ws.column_dimensions['O'].width = 25 # 위험성감소대책 expander
ws.column_dimensions['P'].width = 7
ws.column_dimensions['Q'].width = 7
ws.column_dimensions['R'].width = 7

# --- Data Definition ---
data = [
    # 1. 사전준비
    ["사전준비", "기계적\n(설비)\n요인", "필름부착 작업 시 배관상부에서 미끄러져 넘어짐 위험", "안전모, 안전화 착용 / 작업 전 점검", 3, 2, "V", "1. 2m 이상 고소작업 시 안전대(그네식) 지정 생명줄에 체결 필수\n2. 우천/결빙 시 배관 상부 탑승 작업 전면 중지"],
    ["사전준비", "기계적\n(설비)\n요인", "지하 작업 중 낙하물에 의한 사고위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 굴착 상단부에 자재 및 공구 적재 금지\n2. 하부 작업 시 상부 타 공정 혼재 작업 원칙적 금지"],
    ["사전준비", "기계적\n(설비)\n요인", "필름부착작업 시 사다리 이동작업 중 추락/전도 위험", "안전모, 안전화 착용 / 작업 전 점검", 3, 2, "V", "1. A형 사다리 사용 시 반드시 2인 1조 작업(1인 지지)\n2. 사다리 아웃트리거 전개 및 최상단 발판 탑승 절대 금지"],
    ["사전준비", "기계적\n(설비)\n요인", "차폐체 설치 작업 시 낙하에 의한 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 100kg 이상 중량물 인양 시 2줄 걸이 이상 결속\n2. 인양 중인 중량물 하부 통행 전면 금지"],
    ["사전준비", "기계적\n(설비)\n요인", "차폐체설치 체인블럭 고정상태 불량시 낙하에 의한 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 인양 전 체인블록, 훅(Hook) 마모/파손 상태 사전 점검\n2. 고정 지지대 하중 강도 확인 후 거치"],
    ["사전준비", "기계적\n(설비)\n요인", "손상된 슬링벨트 사용 시 판단에 의한 중량물낙하 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 작업 전 슬링벨트 손상 및 규격 확인\n2. 파손된 슬링벨트는 현장에서 즉시 파기 및 교체"],
    ["사전준비", "기계적\n(설비)\n요인", "중량물 줄걸이작업 부적합 시 낙하에 의한 사고위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 지정된 줄걸이 용구 사용 및 결속 상태 교차 점검\n2. 관리감독자(또는 신호수) 지휘 하에 안전하게 인양 진행"],
    ["사전준비", "기계적\n(설비)\n요인", "근로자가 안전모 등 개인보호구 미착용에 의한 충돌", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 작업구역 진입 전 관리감독자의 개인보호구 착용 상태 점검\n2. 안전모 턱끈 결속 철저"],
    ["사전준비", "기계적\n(설비)\n요인", "중장비 이용 중량물 취급 작업 중 신호오류에 의한 충돌사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 장비 작업 반경 내 접근통제 펜스 설치 및 출입 금지\n2. 전담 신호수 배치 및 운전원 간 무전 소통 확립"],
    ["사전준비", "기계적\n(설비)\n요인", "필름부착 작업 시 웰딩하우스 하부 이동 중 충돌위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 웰딩하우스 하단 돌출 부위에 야광 완충재(스펀지) 부착\n2. 조도(밝기) 확보 및 보행 시 전방 주시 교육"],

    # 2. 비파괴 검사
    ["비파괴\n검사", "화학(물질)적\n요인", "필름현상작업 시 약품취급오류에 의한 사고 위험", "MSDS 준수 보호구 착용", 1, 1, "VI", "1. 화학물질 취급 전 특별 교육 실시\n2. 작업 시 내화학 장갑 및 방독마스크 착용"],
    ["비파괴\n검사", "화학(물질)적\n요인", "필름현상작업 시 취급약품 누출에 의한 사고 위험", "MSDS 준수 보호구 착용", 1, 1, "VI", "1. 약품 보관 용기 밀폐 및 별도 보관소 운영\n2. 암실 및 현상 구역 내 환기 장치 상시 가동"],
    ["비파괴\n검사", "화학(물질)적\n요인", "대상 화학물질에 대한 유해위험성 미인식에 의한 건강장해 발생 위험", "MSDS 준수 보호구 착용", 1, 1, "VI", "1. 현상액/정착액 MSDS(물질안전보건자료) 현장 내 상시 비치\n2. 누출 사고 발생 시 응급 대처 방법 TBM 전파"],
    ["비파괴\n검사", "화학(물질)적\n요인", "화학물질 취급작업시 보호구 미착용으로 인한 사고 위험", "MSDS 준수 보호구 착용", 1, 1, "VI", "1. 화학약품 취급 전 개인보호구 착용 상태 상호 확인\n2. 보호구 훼손 시 즉시 새 제품으로 교체 지급"],
    ["비파괴\n검사", "생물학적\n요인", "검사안전구역 미 설정 시 방사능 노출에 의한 건강장해 위험", "작업 전, 중, 후 방사선량을 측정", 1, 3, "VI", "1. 콜리메이터(차폐장비) 장착 및 10μSv/hr 이하 통제구역 설정\n2. 방사선 감시자 배치 및 일반인 출입 원천 차단\n3. 알람메타 항시 패용"],
    ["비파괴\n검사", "작업특성\n요인", "협소공간 필름부착 시 부자연스런 자세에 의한 근골격계질환 위험", "작업 전 스트레칭", 2, 1, "VI", "1. 작업 전·후 전신 스트레칭 실시\n2. 장시간 구부린 자세 작업 시 주기적인 휴식 시간(10~15분) 부여"],
    ["비파괴\n검사", "작업환경\n요인", "우천으로 인한 굴착법면상태 악화에 의한 붕괴 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 작업 전 굴착 법면 균열, 용수(물비침) 현상 일일 점검\n2. 호우 경보 발효 시 굴착 지점 내부 진입 전면 금지"],

    # 3. 작업후 정리
    ["작업후\n정리", "기계적(설비)적\n요인", "필름해체 작업 시 배관상부에서 미끄러져 넘어짐 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 해체 작업 중에도 안전대 체결 상태 지속 유지\n2. 해체된 폐자재는 던지지 않고 포대/로프를 이용해 안전하게 하강"],
    ["작업후\n정리", "기계적(설비)적\n요인", "중량물 줄걸이 작업 부적합 시 낙하에 의한 사고위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 차폐체 등 반출 시에도 2줄 걸이 이상 결속 원칙 준수\n2. 인양물 하부 접근 금지"],
    ["작업후\n정리", "기계적(설비)적\n요인", "손상된 슬링벨트 사용 시 판단에 의한 중량물낙하 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 철수 작업 전 슬링벨트 훼손 유무 재차 확인\n2. 손상 의심 시 무리한 사용 금지 및 교체"],
    ["작업후\n정리", "기계적(설비)적\n요인", "필름해체작업 시 사다리 이동작업 중 추락/전도 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 해체 시에도 A형 사다리 2인 1조 작업 원칙 준수\n2. 아웃트리거 고정 상태 재점검"],
    ["작업후\n정리", "기계적(설비)적\n요인", "차폐체 해체 작업 시 낙하에 의한 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 차폐체 볼트/너트 해체 전 체인블록 장력 유지 확인\n2. 해체물 임의 적재 금지"],
    ["작업후\n정리", "기계적(설비)적\n요인", "필름해체 작업 시 웰딩하우스 하부 이동 중 충돌위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 철수 시 조도 저하에 대비해 휴대용 랜턴 지참\n2. 보행 시 전방 주시 및 안전모 착용 유지"],
    ["작업후\n정리", "기계적(설비)적\n요인", "중장비 이용 중량물 취급작업 중 신호오류에 의한 충돌사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 장비 철수 동선 사전 파악 및 유도자 배치\n2. 운전원과 눈맞춤 및 무전 확인 후 이동"],
    ["작업후\n정리", "작업특성\n요인", "협소공간 필름해체 시 부자연스런 자세에 의한 근골격계질환 위험", "작업 후 스트레칭", 2, 2, "V", "1. 무리한 힘을 가해 필름 및 차폐체를 해체하지 않음\n2. 일과 종료 후 전신 스트레칭으로 근골격계 긴장 완화"]
]

start_row = 13
current_row = 13

groups = {}
for item in data:
    cat = item[0]
    groups[cat] = groups.get(cat, 0) + 1

for item in data:
    ws.cell(row=current_row, column=1, value=item[0]).alignment = center_align
    ws.cell(row=current_row, column=2, value=item[1]).alignment = center_align
    
    ws.merge_cells(f'C{current_row}:E{current_row}')
    ws.cell(row=current_row, column=3, value=item[2]).alignment = left_align
    ws.merge_cells(f'F{current_row}:G{current_row}')
    ws.cell(row=current_row, column=6, value=item[3]).alignment = left_align
    ws.cell(row=current_row, column=8, value=item[4]).alignment = center_align
    ws.cell(row=current_row, column=9, value=item[5]).alignment = center_align
    ws.cell(row=current_row, column=10, value=item[6]).alignment = center_align
    
    ws.merge_cells(f'K{current_row}:O{current_row}')
    ws.cell(row=current_row, column=11, value=item[7]).alignment = left_align
    
    ws.row_dimensions[current_row].height = 40
    current_row += 1

r = 13
for cat, count in groups.items():
    if count > 1:
        ws.merge_cells(f'A{r}:A{r+count-1}')
    r += count

def merge_sub_categories():
    r_idx = 13
    while r_idx < current_row:
        val = ws.cell(row=r_idx, column=2).value
        count = 1
        for j in range(r_idx + 1, current_row):
            if ws.cell(row=j, column=2).value == val and ws.cell(row=j, column=1).value == ws.cell(row=r_idx, column=1).value:
                count += 1
            else:
                break
        if count > 1:
            ws.merge_cells(f'B{r_idx}:B{r_idx+count-1}')
        r_idx += count

merge_sub_categories()

set_border(ws, 1, 13, 21, current_row - 1)

try:
    wb.save(output_path)
    print(f"Excel file created at: {output_path}")
except PermissionError:
    # If the user has it open, save to a new name
    alt_path = output_path.replace(".xlsx", "_비율수정.xlsx")
    wb.save(alt_path)
    print(f"Excel file created at: {alt_path}")
