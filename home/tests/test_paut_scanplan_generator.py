import importlib.util
from pathlib import Path
import tempfile
import sys

from openpyxl import load_workbook


MODULE = Path(__file__).parents[1] / "src" / "paut_scanplan_generator.py"
spec = importlib.util.spec_from_file_location("paut_scanplan_generator", MODULE)
mod = importlib.util.module_from_spec(spec)
assert spec.loader is not None
sys.modules[spec.name] = mod
spec.loader.exec_module(mod)


def test_generate_scanplan_workbook():
    image = Path(r"C:\Users\-\OneDrive\사진\스크린샷\스크린샷 2026-08-12 184805.png")
    if not image.exists():
        return
    with tempfile.TemporaryDirectory() as tmp:
        output = Path(tmp) / "result.xlsx"
        config = mod.ScanPlanConfig(
            image_path=image,
            output_path=output,
            thickness=28.55,
            indexes=(28.0, 56.0, 84.0),
        )
        generated = mod.create_workbook(config)
        mod.verify_workbook(generated, config)
        wb = load_workbook(generated, data_only=False)
        ws = wb["PAUT 설정표"]
        assert ws["B5"].value == 28.55
        assert ws["A14"].value == "5L64-A2"
        assert [ws[c].value for c in ("I9", "J9", "K9")] == [28, 56, 84]
        assert len(ws._images) == 1


if __name__ == "__main__":
    test_generate_scanplan_workbook()
    print("ok")
