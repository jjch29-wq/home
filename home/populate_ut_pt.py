import openpyxl
from openpyxl.cell.cell import MergedCell
import math
import re

def get_dia_val(dia_str):
    if not dia_str: return None
    dia_str = str(dia_str).replace('"', '').replace("'", '').replace('(wed)', '').strip()
    dia_str = dia_str.replace('”', '').replace('“', '')
    if not dia_str: return None
    try:
        # Handle cases like "1 1/2" or "1 1/2\""
        parts = dia_str.split()
        if len(parts) == 2 and '/' in parts[1]:
            whole = float(parts[0])
            num, den = parts[1].split('/')
            return whole + float(num) / float(den)
        if '/' in dia_str:
            num, den = dia_str.split('/')
            return float(num) / float(den)
        return float(dia_str)
    except:
        return None

def calc_length(dia):
    # diameter * 25.4 * 3.14, truncated to 3 decimal places
    val = dia * 25.4 * 3.14
    return math.floor(val * 1000) / 1000.0

def write_val(ws, r, c, val):
    cell = ws.cell(row=r, column=c)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left_cell.value = val
                return
    else:
        cell.value = val

excel_path = r"C:\Users\-\OneDrive\바탕 화면\템플릿 가산가평월간용역진도보고서.xlsx"
wb = openpyxl.load_workbook(excel_path)

sheets_to_update = ["나. 검사물량 세부내역 (주배관)", "나. 검사물량 세부내역 (관리소)"]

for sheet_name in sheets_to_update:
    if sheet_name not in wb.sheetnames: continue
    ws = wb[sheet_name]
    
    in_ut_pt_section = False
    
    for r in range(20, 150):
        val1 = ws.cell(row=r, column=1).value
        
        if isinstance(val1, str) and ("초음파탐상" in val1 or "UT" in val1.upper() or "액체침투탐상" in val1 or "PT" in val1.upper()):
            in_ut_pt_section = True
            continue
            
        if in_ut_pt_section:
            # Check if section ends
            if isinstance(val1, str) and ("총" in val1 or "합" in val1 or "비고" in val1 or "※" in val1):
                in_ut_pt_section = False
                continue
            
            dia_cell = ws.cell(row=r, column=2).value
            dia_val = get_dia_val(dia_cell)
            
            if dia_val is not None:
                length = calc_length(dia_val)
                # Write to Col 3
                write_val(ws, r, 3, length)

wb.save(excel_path)
print("Successfully populated inspection lengths for UT and PT.")
