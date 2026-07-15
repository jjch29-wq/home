import openpyxl

filepath = r'C:\Users\-\OneDrive\바탕 화면\지역난방 PAUT.xlsx'
try:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = next((sn for sn in wb.sheetnames if '을지' in sn or 'DATA' in sn.upper()), wb.sheetnames[0])
    ws = wb[sheet_name]
    
    col_start = 'Y'
    col_end = 'AH'
    s_idx = openpyxl.utils.column_index_from_string(col_start)
    e_idx = openpyxl.utils.column_index_from_string(col_end)
    
    tot = 0.0
    w_list = []
    
    for i in range(s_idx, e_idx + 1):
        let = openpyxl.utils.get_column_letter(i)
        cd = ws.column_dimensions.get(let)
        w = cd.width if cd and cd.width else ws.sheet_format.defaultColWidth
        if w is None:
            w = 8.38
        w_list.append(f"{let}:{w:.2f}")
        tot += w
        
    print(f"Columns Y to AH Widths: {', '.join(w_list)}")
    print(f"Total Width (Y-AH): {tot:.2f}")
except Exception as e:
    print(f"Error: {e}")
