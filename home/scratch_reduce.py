import openpyxl

filepath = r'C:\Users\-\OneDrive\바탕 화면\지역난방 PAUT.xlsx'
try:
    wb = openpyxl.load_workbook(filepath)
    
    sheet_name = None
    for sn in wb.sheetnames:
        if '을지' in sn or 'DATA' in sn.upper() or 'Data' in sn:
            sheet_name = sn
            break
            
    if not sheet_name:
        sheet_name = wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]

    ws = wb[sheet_name]
    
    s_col = 'S'
    cd = ws.column_dimensions.get(s_col)
    current_width = cd.width if cd and cd.width else ws.sheet_format.defaultColWidth
    if current_width is None:
        current_width = 8.38
        
    ws.column_dimensions[s_col].width = current_width - 2.0
    
    # Recalculate
    col_start = 'N'
    col_end = 'X'
    s_idx = openpyxl.utils.column_index_from_string(col_start)
    e_idx = openpyxl.utils.column_index_from_string(col_end)
    
    tot = 0.0
    for i in range(s_idx, e_idx + 1):
        let = openpyxl.utils.get_column_letter(i)
        cd = ws.column_dimensions.get(let)
        w = cd.width if cd and cd.width else ws.sheet_format.defaultColWidth
        if w is None:
            w = 8.38
        tot += w
        
    wb.save(filepath)
    print(f"SUCCESS: Total width is now {tot:.2f}")
except PermissionError:
    print("PERMISSION_ERROR")
except Exception as e:
    print(f"Error: {e}")
