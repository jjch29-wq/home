import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

def create_dorm_guidelines(output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "숙소 생활 지침"

    # A4 세로 인쇄 설정
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    
    # 여백 최소화
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    ws.print_area = 'A1:A16'

    # 열 너비 (A4 가로 폭 꽉 차게)
    ws.column_dimensions['A'].width = 110

    # 행 높이 설정 (총 높이 약 850 내외로 분배)
    ws.row_dimensions[1].height = 90   # 타이틀
    ws.row_dimensions[2].height = 20   # 여백
    ws.row_dimensions[3].height = 45   # 1
    ws.row_dimensions[4].height = 35   # 1-sub
    ws.row_dimensions[5].height = 45   # 2
    ws.row_dimensions[6].height = 45   # 3
    ws.row_dimensions[7].height = 45   # 4
    ws.row_dimensions[8].height = 45   # 5
    ws.row_dimensions[9].height = 35   # 5-sub
    ws.row_dimensions[10].height = 45  # 6
    ws.row_dimensions[11].height = 35  # 6-sub
    ws.row_dimensions[12].height = 45  # 7
    ws.row_dimensions[13].height = 35  # 7-sub
    ws.row_dimensions[14].height = 45  # 8
    ws.row_dimensions[15].height = 30  # 여백
    ws.row_dimensions[16].height = 80  # Footer

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thick = Side(style='thick', color='000000')
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)
    border_left_right = Border(left=thick, right=thick)
    
    def set_cell(row, val, size, color='000000', bold=True, bg_color=None, align=align_left, border=border_left_right):
        c = ws.cell(row=row, column=1)
        c.value = val
        c.font = Font(name='맑은 고딕', size=size, bold=bold, color=color)
        c.alignment = align
        c.border = border
        if bg_color:
            c.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        return c

    # 상단 테두리를 위해 A1은 특별히 위쪽 테두리 추가
    c1 = set_cell(1, "☆☆ 숙소 생활 지침 ☆☆", 45, color='FFFFFF', bg_color='4472C4', align=align_center)
    c1.border = Border(left=thick, right=thick, top=thick, bottom=thick)
    
    set_cell(2, "", 20) # 여백
    
    set_cell(3, "  *. 실내 흡연 금지 (임대조건)", 28, color='FF0000')
    set_cell(4, "      - 흡연 냄새 발생시 계약종료 때 벽지 재시공 조건", 22, color='595959')
    
    set_cell(5, "  *. 벽에 못 시공 금지 (임대조건)", 28)
    
    set_cell(6, "  *. 입주 전 하자 부분 채증(사진) 확보", 28)
    
    set_cell(7, "  *. 입주 전 전기, 가스 계량기 사진 확보", 28)
    
    set_cell(8, "  *. 하자 발생시 원인 제공자 즉시 보고 및 복원", 28)
    set_cell(9, "      - 원인 제공자 확인서 제출 / 비용 청구 근거", 22, color='595959')
    
    set_cell(10, "  *. 실내 청결 유지 (공동, 개별 사용 구역 청소)", 28)
    set_cell(11, "      - 음식물 쓰레기 발생 즉시 / 생활 쓰레기 수시 처리", 22, color='595959')
    
    set_cell(12, "  *. 가스, 전기료 회사 상한선 초과시 공동 분배", 28)
    set_cell(13, "      - 냉, 난방기 관리 철저 (야간조 편성시 특히 주의)", 22, color='595959')
    
    set_cell(14, "  *. 침구류 개별 구매 관리 - 사무소 제공 無 (지침)", 28)
    
    set_cell(15, "", 20) # 여백
    
    # 하단 꼬리말 부분은 둥근 테두리 느낌을 위해 아래쪽 테두리 추가
    c16 = set_cell(16, "☆☆ 상기 사항 유지 관리 수시 점검 예정 ☆☆", 30, color='000000', bg_color='FFF2CC', align=align_center)
    c16.border = Border(left=thick, right=thick, top=thick, bottom=thick)

    # 파일 저장
    wb.save(output_path)
    print(f"안내문이 바탕화면에 저장되었습니다: {output_path}")

if __name__ == "__main__":
    desktop = os.path.join(os.path.expanduser("~"), "OneDrive", "바탕 화면")
    if not os.path.exists(desktop):
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        
    output_file = os.path.join(desktop, "숙소_생활_지침_안내문.xlsx")
    create_dorm_guidelines(output_file)
