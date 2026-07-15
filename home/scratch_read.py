import openpyxl

filepath = r'C:\Users\-\OneDrive\바탕 화면\지역난방 PAUT.xlsx'
try:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = next((sn for sn in wb.sheetnames if '을지' in sn or 'DATA' in sn.upper()), wb.sheetnames[0])
    ws = wb[sheet_name]
    
    print(f"Max Col: {ws.max_column}")
    
    for row in ws.iter_rows(min_col=24, max_col=35, min_row=1, max_row=15, values_only=True):
        print(row)
        
except Exception as e:
    print(f"Error: {e}")
