import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "작업현장 현황"

# Title
ws.merge_cells('A1:G1')
ws['A1'] = "8월 광복절 연휴 작업현장 현황"
ws['A1'].font = Font(size=18, bold=True)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Headers Row 2
ws.merge_cells('A2:A3')
ws['A2'] = "구분"
ws.merge_cells('B2:B3')
ws['B2'] = "공사명"

ws.merge_cells('C2:E2')
ws['C2'] = "작업계획"
ws['C3'] = "8.15 (토) 광복절"
ws['D3'] = "8.16 (일)"
ws['E3'] = "8.17 (월) 대체휴일"

ws.merge_cells('F2:F3')
ws['F2'] = "공사감독 등 입회\n(공사관리관)"
ws.merge_cells('G2:G3')
ws['G2'] = "특이사항"

# Data Row 4
ws['A4'] = "건설사업단\n(서울건설)"
ws['B4'] = "가산~가평 천연가스\n공급시설 건설공사"
ws['C4'] = "정상작업"
ws['D4'] = "정상작업"
ws['E4'] = "휴무"
ws['F4'] = "(8.15) 차장 OOO\n(8.16) 차장 OOO\n(8.17) 차장 OOO"
ws['G4'] = ""

# Styling
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row=4, min_col=1, max_col=7):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if cell.row in [2, 3]:
            cell.fill = header_fill
            cell.font = Font(bold=True)

# Set text colors for dates (as they appear reddish in the image)
ws['C3'].font = Font(color="C00000", bold=True)
ws['D3'].font = Font(color="C00000", bold=True)
ws['E3'].font = Font(color="C00000", bold=True)

# Adjust column widths and row heights
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 15

ws.row_dimensions[1].height = 40
ws.row_dimensions[4].height = 60

# Save
wb.save(r"c:\Users\-\PMI\home\8월_광복절_연휴_작업현장_현황.xlsx")
print("Excel file created.")
