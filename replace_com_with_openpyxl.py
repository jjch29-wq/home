import codecs
import re

file_path = r'c:\Users\jjch2\Desktop\PMI\home\src\한국지역난방 중앙지사.py'
with codecs.open(file_path, 'r', 'utf-8') as f:
    text = f.read()

# I will replace the entire "--- 5. 엑셀 기입 ---" block with an openpyxl version!
# To do this safely, I will write the replacement logic.

openpyxl_code = '''                # --- 5. 엑셀 기입 (openpyxl로 변경) ---
                import openpyxl
                from openpyxl.utils import get_column_letter
                import os
                
                wb = openpyxl.load_workbook(save_path)
                
                def write_ndt_sheet_openpyxl(ws, agg):
                    if agg is None: return
                    
                    def get_val(bucket, shift, insp_type, field='qty'):
                        key = (shift, insp_type)
                        return bucket.get(key, {}).get(field, 0)
                    
                    def safe_set(row, col, val):
                        if val and val != 0:
                            ws.cell(row=row, column=col).value = round(val, 2) if isinstance(val, float) and val != int(val) else int(val) if isinstance(val, float) and val == int(val) else val
                            
                    rt_types = {'B': 8, 'A': 9, 'A/2': 10}
                    rt_total = {'주간_joints': 0, '주간_ORI': 0, '주간_REP': 0,
                                '야간_joints': 0, '야간_ORI': 0, '야간_REP': 0}
                    
                    for film_key, row_num in rt_types.items():
                        bucket = agg['RT'][film_key]
                        j_day = get_val(bucket, '주간', 'ORI', 'joints') + get_val(bucket, '주간', 'REP', 'joints')
                        ori_day = get_val(bucket, '주간', 'ORI', 'qty')
                        rep_day = get_val(bucket, '주간', 'REP', 'qty')
                        sum_day = ori_day + rep_day
                        j_night = get_val(bucket, '야간/휴일', 'ORI', 'joints') + get_val(bucket, '야간/휴일', 'REP', 'joints')
                        ori_night = get_val(bucket, '야간/휴일', 'ORI', 'qty')
                        rep_night = get_val(bucket, '야간/휴일', 'REP', 'qty')
                        sum_night = ori_night + rep_night
                        j_total = j_day + j_night
                        ori_total = ori_day + ori_night
                        rep_total = rep_day + rep_night
                        sum_total = ori_total + rep_total
                        
                        safe_set(row_num, 3, j_day)
                        safe_set(row_num, 4, ori_day)
                        safe_set(row_num, 5, rep_day)
                        safe_set(row_num, 6, sum_day)
                        safe_set(row_num, 7, j_night)
                        safe_set(row_num, 8, ori_night)
                        safe_set(row_num, 9, rep_night)
                        safe_set(row_num, 10, sum_night)
                        safe_set(row_num, 11, j_total)
                        safe_set(row_num, 12, ori_total)
                        safe_set(row_num, 13, rep_total)
                        safe_set(row_num, 14, sum_total)
                        
                        rt_total['주간_joints'] += j_day
                        rt_total['주간_ORI'] += ori_day
                        rt_total['주간_REP'] += rep_day
                        rt_total['야간_joints'] += j_night
                        rt_total['야간_ORI'] += ori_night
                        rt_total['야간_REP'] += rep_night
                        
                    safe_set(11, 3, rt_total['주간_joints'])
                    safe_set(11, 4, rt_total['주간_ORI'])
                    safe_set(11, 5, rt_total['주간_REP'])
                    safe_set(11, 6, rt_total['주간_ORI'] + rt_total['주간_REP'])
                    safe_set(11, 7, rt_total['야간_joints'])
                    safe_set(11, 8, rt_total['야간_ORI'])
                    safe_set(11, 9, rt_total['야간_REP'])
                    safe_set(11, 10, rt_total['야간_ORI'] + rt_total['야간_REP'])
                    safe_set(11, 11, rt_total['주간_joints'] + rt_total['야간_joints'])
                    safe_set(11, 12, rt_total['주간_ORI'] + rt_total['야간_ORI'])
                    safe_set(11, 13, rt_total['주간_REP'] + rt_total['야간_REP'])
                    safe_set(11, 14, rt_total['주간_ORI'] + rt_total['주간_REP'] + rt_total['야간_ORI'] + rt_total['야간_REP'])
                    
                    ut = agg['UT']['data']
                    ut_day_j = get_val(ut, '주간', 'ORI', 'joints') + get_val(ut, '주간', 'REP', 'joints')
                    ut_day_ori = get_val(ut, '주간', 'ORI', 'qty')
                    ut_day_rep = get_val(ut, '주간', 'REP', 'qty')
                    ut_night_j = get_val(ut, '야간/휴일', 'ORI', 'joints') + get_val(ut, '야간/휴일', 'REP', 'joints')
                    ut_night_ori = get_val(ut, '야간/휴일', 'ORI', 'qty')
                    ut_night_rep = get_val(ut, '야간/휴일', 'REP', 'qty')
                    
                    safe_set(12, 3, ut_day_j)
                    safe_set(12, 4, ut_day_ori + ut_day_rep)
                    safe_set(12, 6, ut_day_ori + ut_day_rep)
                    safe_set(12, 7, ut_night_j)
                    safe_set(12, 8, ut_night_ori + ut_night_rep)
                    safe_set(12, 10, ut_night_ori + ut_night_rep)
                    safe_set(12, 11, ut_day_j + ut_night_j)
                    safe_set(12, 12, ut_day_ori + ut_day_rep + ut_night_ori + ut_night_rep)
                    safe_set(12, 14, ut_day_ori + ut_day_rep + ut_night_ori + ut_night_rep)
                    
                    ut_day_adj = get_val(ut, '주간', 'ORI', 'adj_qty') + get_val(ut, '주간', 'REP', 'adj_qty')
                    ut_night_adj = get_val(ut, '야간/휴일', 'ORI', 'adj_qty') + get_val(ut, '야간/휴일', 'REP', 'adj_qty')
                    safe_set(13, 4, ut_day_adj)
                    safe_set(13, 6, ut_day_adj)
                    safe_set(13, 8, ut_night_adj)
                    safe_set(13, 10, ut_night_adj)
                    safe_set(13, 12, ut_day_adj + ut_night_adj)
                    safe_set(13, 14, ut_day_adj + ut_night_adj)
                    
                    pt = agg['PT']['data']
                    pt_day_j = get_val(pt, '주간', 'ORI', 'joints') + get_val(pt, '주간', 'REP', 'joints')
                    pt_day_ori = get_val(pt, '주간', 'ORI', 'qty')
                    pt_day_rep = get_val(pt, '주간', 'REP', 'qty')
                    pt_night_j = get_val(pt, '야간/휴일', 'ORI', 'joints') + get_val(pt, '야간/휴일', 'REP', 'joints')
                    pt_night_ori = get_val(pt, '야간/휴일', 'ORI', 'qty')
                    pt_night_rep = get_val(pt, '야간/휴일', 'REP', 'qty')
                    
                    safe_set(14, 3, pt_day_j)
                    safe_set(14, 4, pt_day_ori + pt_day_rep)
                    safe_set(14, 6, pt_day_ori + pt_day_rep)
                    safe_set(14, 7, pt_night_j)
                    safe_set(14, 8, pt_night_ori + pt_night_rep)
                    safe_set(14, 10, pt_night_ori + pt_night_rep)
                    safe_set(14, 11, pt_day_j + pt_night_j)
                    safe_set(14, 12, pt_day_ori + pt_day_rep + pt_night_ori + pt_night_rep)
                    safe_set(14, 14, pt_day_ori + pt_day_rep + pt_night_ori + pt_night_rep)
                    
                    pt_day_adj = get_val(pt, '주간', 'ORI', 'adj_qty') + get_val(pt, '주간', 'REP', 'adj_qty')
                    pt_night_adj = get_val(pt, '야간/휴일', 'ORI', 'adj_qty') + get_val(pt, '야간/휴일', 'REP', 'adj_qty')
                    safe_set(15, 4, pt_day_adj)
                    safe_set(15, 6, pt_day_adj)
                    safe_set(15, 8, pt_night_adj)
                    safe_set(15, 10, pt_night_adj)
                    safe_set(15, 12, pt_day_adj + pt_night_adj)
                    safe_set(15, 14, pt_day_adj + pt_night_adj)
                    
                    total_film = rt_total['주간_ORI'] + rt_total['주간_REP'] + rt_total['야간_ORI'] + rt_total['야간_REP']
                    safe_set(16, 11, 0)
                    safe_set(16, 12, total_film)
                    
                sheet_names = wb.sheetnames
                if main_agg:
                    sheet_name = '3. 비파괴검사 현황 (열배관)'
                    if sheet_name in sheet_names:
                        write_ndt_sheet_openpyxl(wb[sheet_name], main_agg)
                        log(f"✅ '{sheet_name}' 시트 기입 완료")
                    else:
                        log(f"⚠️ '{sheet_name}' 시트를 찾을 수 없습니다.")
                        
                if mgmt_agg:
                    sheet_name = '3. 비파괴검사 현황 (관리소)'
                    if sheet_name in sheet_names:
                        write_ndt_sheet_openpyxl(wb[sheet_name], mgmt_agg)
                        log(f"✅ '{sheet_name}' 시트 기입 완료")
                    else:
                        log(f"⚠️ '{sheet_name}' 시트를 찾을 수 없습니다.")
                        
                paut_sheet_name = '3. 비파괴검사 현황 (열배관)'
                if paut_sheet_name in sheet_names:
                    try:
                        paut_records = []
                        target_month_str = f"{year}-{month:02d}"
                        
                        for date_key, log_data in history.items():
                            if date_key.startswith(target_month_str):
                                ndt_results = log_data.get('ndt_results', [])
                                for r in ndt_results:
                                    if str(r.get('검사방법', '')).strip().upper() == 'PAUT':
                                        r['_date'] = date_key
                                        paut_records.append(r)
                                        
                        if paut_records:
                            paut_records.sort(key=lambda x: x['_date'])
                            
                            groups = {}
                            for r in paut_records:
                                key = (str(r.get('업체', '')), str(r.get('구간', '')), str(r.get('라인번호', '')), str(r.get('관경', '')), str(r.get('Joint No.', '')))
                                paut_val = str(r.get('PAUT', '0')).strip()
                                try: val = float(paut_val)
                                except: val = 0.0
                                if val == 0: continue
                                shift = str(r.get('규격', '주간')).strip()
                                
                                if key not in groups:
                                    groups[key] = {'ORI': 0.0, 'RE': 0.0, 'shift': shift}
                                if groups[key]['ORI'] == 0.0:
                                    groups[key]['ORI'] += val
                                else:
                                    groups[key]['RE'] += val
                                    
                            ws_paut = wb[paut_sheet_name]
                            
                            if len(groups) > 1:
                                ws_paut.insert_rows(406, amount=len(groups)-1)
                                
                            current_row = 405
                            for i, (key, data) in enumerate(groups.items()):
                                업체, 구간, 라인번호, 관경, joint = key
                                ws_paut.cell(row=current_row, column=2).value = 업체
                                ws_paut.cell(row=current_row, column=3).value = i + 1
                                ws_paut.cell(row=current_row, column=4).value = 구간
                                ws_paut.cell(row=current_row, column=5).value = 라인번호
                                ws_paut.cell(row=current_row, column=6).value = 관경
                                ws_paut.cell(row=current_row, column=7).value = joint
                                ws_paut.cell(row=current_row, column=8).value = data['shift']
                                ws_paut.cell(row=current_row, column=9).value = 'M'
                                
                                ori = data['ORI']
                                re_val = data['RE']
                                tot = ori + re_val
                                
                                if ori > 0: ws_paut.cell(row=current_row, column=10).value = round(ori, 4)
                                if re_val > 0: ws_paut.cell(row=current_row, column=11).value = round(re_val, 4)
                                if tot > 0: ws_paut.cell(row=current_row, column=12).value = round(tot, 4)
                                
                                current_row += 1
                                
                            log(f"▶ '{paut_sheet_name}' 시트 기입 완료 (총 {len(groups)}건)")
                    except Exception as e:
                        log(f"⚠️ '{paut_sheet_name}' 작성 중 오류: {e}")
                        
                wb.save(save_path)
                wb.close()
                log(f"\\n🎉 저장 완료: {save_path}")
                messagebox.showinfo("완료", f"월간 진도보고서 비파괴검사 현황이 업데이트되었습니다.\\n{save_path}")
                import os
                os.startfile(os.path.dirname(save_path))
                
            except Exception as e:
'''

start_idx = text.find('# --- 5. 엑셀 기입 ---')
end_idx = text.find('except Exception as e:', start_idx)

if start_idx != -1 and end_idx != -1:
    new_text = text[:start_idx] + openpyxl_code + text[end_idx + 23:]
    with codecs.open(file_path, 'w', 'utf-8') as f:
        f.write(new_text)
    print("Replaced COM with openpyxl.")
else:
    print("Could not find the block to replace!")
