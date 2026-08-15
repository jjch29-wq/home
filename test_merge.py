import openpyxl
import glob

template_file = glob.glob(r'c:\Users\jjch2\Desktop\*지역난방*RT*.xlsx')[0]
wb = openpyxl.load_workbook(template_file)
ws0 = wb.worksheets[0]

print("Merged cells before:")
for merge in ws0.merged_cells.ranges:
    if 'H' in str(merge) or 'O' in str(merge):
        print("  ", merge)

try:
    ws0.unmerge_cells('H1:O3')
    print("Unmerge H1:O3 successful")
except Exception as e:
    print("Unmerge error:", e)

try:
    ws0.merge_cells('H1:O4')
    print("Merge H1:O4 successful")
except Exception as e:
    print("Merge error:", e)

print("Merged cells after:")
for merge in ws0.merged_cells.ranges:
    if 'H' in str(merge) or 'O' in str(merge):
        print("  ", merge)
