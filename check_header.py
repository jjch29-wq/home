import openpyxl, glob, os
from openpyxl.utils import get_column_letter

files = [f for f in glob.glob(r'C:\Users\jjch2\Desktop\*.xlsx') if not os.path.basename(f).startswith('~$')]
files.sort(key=os.path.getmtime, reverse=True)
wb = openpyxl.load_workbook(files[0])
ws = wb.worksheets[0]

# Find title merges (F?:O?) and corresponding doc merges (P?:W?)
title_merges = []
doc_merges = []

for m in ws.merged_cells.ranges:
    val = str(ws.cell(m.min_row, m.min_col).value or '')
    # Title merges typically start at col F(6) and end at col O(15)
    if m.min_col == 6 and m.max_col == 15:
        title_merges.append(m)
    # Doc merges typically start at col P(16) and end at col W(23)  
    if m.min_col == 16 and m.max_col == 23:
        doc_merges.append(m)

print("Title merges (F:O):")
for m in sorted(title_merges, key=lambda x: x.min_row):
    print(f"  {m} -> rows {m.min_row}-{m.max_row}")

print(f"\nDoc merges (P:W):")
for m in sorted(doc_merges, key=lambda x: x.min_row):
    val = str(ws.cell(m.min_row, m.min_col).value or '').replace('\n','|')[:40]
    print(f"  {m} -> rows {m.min_row}-{m.max_row}: {val}")

# Also check page 1 (no merges, just cell positions)
print("\nPage 1 header (no merges):")
for row in range(1, 6):
    for col in range(1, 24):
        c = ws.cell(row=row, column=col)
        if c.value:
            val = str(c.value).replace('\n','|')[:40]
            print(f"  {get_column_letter(col)}{row} (col {col}): {val}")
