import openpyxl
import shutil
import re

src_file = r"c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\Na-aba\home\data\가스공사 의뢰서.xlsx"
dest_file = r"c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\scratch\가스공사_의뢰서_테스트.xlsx"
shutil.copy(src_file, dest_file)

wb = openpyxl.load_workbook(dest_file)
ws = wb.worksheets[0]

# Mimic items for Weld 63 at row 14
items = [
    {
        'Loc': '1-2',
        'Deg': '1',
        'D5': 'v', # Porosity
        'Result': 'ACC',
        'Remarks': 'Test1'
    },
    {
        'Loc': '2-3',
        'Deg': '2',
        'Result': 'ACC',
        'Remarks': 'Test2'
    }
]

# Set method requests on row 14: RT = 'O', UT = None, PT = 'O'
ws.cell(row=14, column=13, value='O')
ws.cell(row=14, column=14, value=None)
ws.cell(row=14, column=15, value='O')

row_idx = 14

# Determine final result and remarks
final_res = "ACC"
all_remarks = []
for it in items:
    res = str(it.get('Result', '')).upper()
    if any(x in res for x in ["REJ", "NG", "불합격"]): final_res = "REJ"
    rem = str(it.get('Remarks', '')).strip()
    if rem and rem not in all_remarks: all_remarks.append(rem)

# Write remarks to column 11
ws.cell(row=row_idx, column=11, value=", ".join(all_remarks) if all_remarks else None)

rt_val = ws.cell(row=row_idx, column=13).value
ut_val = ws.cell(row=row_idx, column=14).value
pt_val = ws.cell(row=row_idx, column=15).value

def is_marked(v):
    if v is None: return False
    cleaned = str(v).strip().upper()
    return cleaned in ["O", "V", "X", "●", "합격", "1", "TRUE"]

rt_req = is_marked(rt_val)
ut_req = is_marked(ut_val)
pt_req = is_marked(pt_val)

has_rt_films = False
for it in items:
    loc_str = str(it.get('Loc', '')).strip()
    if re.match(r'^\d+', loc_str):
        has_rt_films = True
        break

defect_map = {
    "D1": "C", "D2": "IP", "D3": "LF", "D4": "S", "D5": "P",
    "D6": "UC", "D7": "RUC", "D8": "BT", "D9": "TI", "D10": "CP",
    "D11": "RC", "D12": "Mis", "D13": "EP", "D14": "SD", "D15": "Oth"
}

def get_item_val(it):
    grade = str(it.get('Deg', '')).strip()
    checked = []
    for d_key, abbrev in defect_map.items():
        val = str(it.get(d_key, '')).strip()
        if val in ["√", "1", "v", "V", "o", "O", "●", "?"]:
            checked.append(abbrev)
    if checked:
        return f"{grade}/{','.join(checked)}" if grade else f"/{','.join(checked)}"
    else:
        return grade if grade else ("1" if any(x in str(it.get('Result', '')).upper() for x in ["ACC", "OK"]) else "4")

if rt_req or has_rt_films:
    for col in range(17, 24):
        ws.cell(row=row_idx, column=col, value=None)
    for it in items:
        loc_str = str(it.get('Loc', '')).strip()
        film_num = None
        match = re.match(r'^(\d+)', loc_str)
        if match:
            film_num = int(match.group(1))
        if film_num and 1 <= film_num <= 7:
            val = get_item_val(it)
            ws.cell(row=row_idx, column=16 + film_num, value=val)
            
if ut_req:
    it = items[0] if items else {}
    val = get_item_val(it)
    ws.cell(row=row_idx, column=24, value=val)
    
if pt_req:
    it = items[0] if items else {}
    val = get_item_val(it)
    ws.cell(row=row_idx, column=25, value=val)

wb.save(dest_file)

# Read it back and verify
wb2 = openpyxl.load_workbook(dest_file, data_only=True)
ws2 = wb2.worksheets[0]

print("Verification of row 14 after writeback:")
print(f"K14 (Remarks) = {ws2.cell(row=14, column=11).value}")
print(f"Q14 (RT Film 1) = {ws2.cell(row=14, column=17).value}")
print(f"R14 (RT Film 2) = {ws2.cell(row=14, column=18).value}")
print(f"X14 (UT/UP) = {ws2.cell(row=14, column=24).value}")
print(f"Y14 (PT) = {ws2.cell(row=14, column=25).value}")
