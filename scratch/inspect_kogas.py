import pandas as pd
import openpyxl

file_path = r"c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\Na-aba\home\data\가스공사 의뢰서.xlsx"

wb = openpyxl.load_workbook(file_path)
print("Sheet names:", wb.sheetnames)
ws = wb.active
print("Active sheet title:", ws.title)

# Print cells from row 9 to 25, columns 1 to 28
print("Cells in row 9 to 16:")
for r in range(9, 17):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 29)]
    print(f"Row {r}: {row_vals}")

# Let's check merged ranges
print("Merged ranges:")
for merged_range in list(ws.merged_cells.ranges):
    print(merged_range)
