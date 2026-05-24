import openpyxl

file_path = r"c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\Na-aba\home\data\SISGI-KOGAS-RT-001.xlsx"

wb = openpyxl.load_workbook(file_path)
print("Sheet names:", wb.sheetnames)
ws = wb.active
print("Active sheet title:", ws.title)

# Print row 9 to 15, first 30 columns
for r in range(9, 16):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 31)]
    print(f"Row {r}: {row_vals}")
