import openpyxl

file_path = r"c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\Na-aba\home\data\SISGI-KOGAS-RT-001.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

for r in range(1, ws.max_row + 1):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 20)]
    # Check if there is any non-empty value in the row
    if any(v is not None for v in vals):
        # If it looks like header or joint data (e.g. contains numbers)
        if any(isinstance(v, int) or (isinstance(v, str) and ('Joint' in v or 'JOINT' in v or 'No' in v or 'NO' in v)) for v in vals):
            print(f"Row {r}: {vals[:15]}")
