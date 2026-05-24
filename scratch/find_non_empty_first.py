import openpyxl

file_path = r"c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\Na-aba\home\data\SISGI-KOGAS-RT-001.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.worksheets[0]

print("Non-empty rows in FIRST sheet:")
for r in range(1, ws.max_row + 1):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 20)]
    if any(v is not None for v in vals):
        print(f"Row {r}: {vals[:15]}")
