import openpyxl

file_path = r"c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\Na-aba\home\data\SISGI-KOGAS-RT-001.xlsx"
wb = openpyxl.load_workbook(file_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Sheet: {sheet_name}, Max Row: {ws.max_row}")
    # print first 5 rows with numbers
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        # If column 2 contains 'Joint No' or 'Welder'
        vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
        if any(isinstance(val, str) and any(x in val.lower() for x in ['joint', 'welder', 'film ident']) for val in vals):
            print(f"  Row {r}: {vals}")
