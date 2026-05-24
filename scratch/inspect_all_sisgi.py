import openpyxl

file_path = r"c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\Na-aba\home\data\SISGI-KOGAS-RT-001.xlsx"
wb = openpyxl.load_workbook(file_path)

for sheet_idx, ws in enumerate(wb.worksheets):
    print(f"\n--- Sheet {sheet_idx}: {ws.title} ---")
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 35)]
        if any(v is not None for v in vals):
            # Print row if it has any non-empty cells
            # Only print first 20 cols
            cleaned_vals = [str(v)[:30] if v is not None else "" for v in vals[:20]]
            print(f"Row {r:02d}: {cleaned_vals}")
