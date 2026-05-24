import openpyxl
import re

defect_map = {
    "D1": "C", "D2": "IP", "D3": "LF", "D4": "S", "D5": "P",
    "D6": "UC", "D7": "RUC", "D8": "BT", "D9": "TI", "D10": "CP",
    "D11": "RC", "D12": "Mis", "D13": "EP", "D14": "SD", "D15": "Oth"
}

def get_item_val(it):
    grade = str(it.get('Deg', '')).strip()
    checked = []
    for d_key, abbrev in defect_map.items():
        val = str(it.get(d_key, '')).strip()
        if val in ["√", "1", "v", "V", "o", "O", "●", "?"]:
            checked.append(abbrev)
    if checked:
        return f"{grade}/{','.join(checked)}" if grade else f"/{','.join(checked)}"
    else:
        return grade if grade else ("1" if any(x in str(it.get('Result', '')).upper() for x in ["ACC", "OK"]) else "4")

def is_marked(v):
    if v is None: return False
    cleaned = str(v).strip().upper()
    return cleaned in ["O", "V", "X", "●", "합격", "1", "TRUE"]

# Simulate KOGAS sync
target_file = r"c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\Na-aba\home\data\가스공사 의뢰서.xlsx"
wb = openpyxl.load_workbook(target_file)
ws = wb.active

# Let's say we have dummy items for row 14
row_idx = 14
items = [
    {
        'Loc': '1-2',
        'Deg': '1',
        'D5': '√',  # P (Porosity)
        'Result': 'ACC',
        '_src': {'sheet': '6월5일-1', 'row': 14}
    },
    {
        'Loc': '2-3',
        'Deg': '2',
        'Result': 'ACC',
        '_src': {'sheet': '6월5일-1', 'row': 14}
    }
]

print("rt_val before:", ws.cell(row=row_idx, column=13).value)
rt_val = ws.cell(row=row_idx, column=13).value
ut_val = ws.cell(row=row_idx, column=14).value
pt_val = ws.cell(row=row_idx, column=15).value

rt_req = is_marked(rt_val)
ut_req = is_marked(ut_val)
pt_req = is_marked(pt_val)

print(f"rt_req: {rt_req}, ut_req: {ut_req}, pt_req: {pt_req}")

has_rt_films = True # force True

if rt_req or has_rt_films:
    print("Clearing Q14:W14 and Q15:W15...")
    for col in range(17, 24):
        ws.cell(row=row_idx, column=col, value=None)
        # Wait, if row 15 is also merged, do we clear row 15 too? Let's check.
    for it in items:
        loc_str = str(it.get('Loc', '')).strip()
        film_num = None
        match = re.match(r'^(\d+)', loc_str)
        if match:
            film_num = int(match.group(1))
        print(f"loc_str: {loc_str}, film_num: {film_num}")
        if film_num and 1 <= film_num <= 7:
            val = get_item_val(it)
            print(f"Writing to col {16+film_num}: {val}")
            ws.cell(row=row_idx, column=16 + film_num, value=val)

wb.save(r"c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\scratch\test_kogas_sync.xlsx")
print("Done saving!")
