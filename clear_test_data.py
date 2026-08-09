import openpyxl
import shutil

# First backup the current file
src = r'C:\Users\jjch2\Desktop\템플릿_최종완성본_V74.xlsx'
bak = r'C:\Users\jjch2\Desktop\템플릿_최종완성본_V74_backup.xlsx'
shutil.copy2(src, bak)
print(f"Backed up to: {bak}")

wb = openpyxl.load_workbook(src)
ws = wb.worksheets[0]

# Clear data rows - keep only header rows (403~404) and structure
# Data area: rows 405 to 480 (before TOTAL row at 481)
# Only clear cells that are NOT in header rows and NOT merged anchors for structure

def is_structural_merge(ws, row, col):
    """해당 셀이 구조적 병합(헤더/테이블 테두리)의 앵커인지 확인"""
    for merge in ws.merged_cells.ranges:
        if merge.min_row == row and merge.min_col == col:
            # 큰 병합 범위면 구조적인 것일 가능성 높음 (width > 5 or height > 2)
            if (merge.max_col - merge.min_col) > 4 or (merge.max_row - merge.min_row) > 1:
                return True
    return False

cleared = 0
# Clear rows 405~480 (data area for 1.2.1 PAUT)
for row in range(405, 481):
    for col in range(1, 30):
        cell = ws.cell(row=row, column=col)
        # Don't clear structural cells
        if cell.value is not None and not is_structural_merge(ws, row, col):
            cell.value = None
            cleared += 1

print(f"Cleared {cleared} cells in rows 405-480")

# Also check if 448~449 is a second header - based on user's info
# Check what rows 448-449 actually have in merged structure
print("\n=== Row 448~449 병합 범위 (헤더 확인) ===")
for merge in ws.merged_cells.ranges:
    if 448 <= merge.min_row <= 449:
        print(f"  {merge.coord}")

wb.save(src)
print(f"\nCleared and saved: {src}")
