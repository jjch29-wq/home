import openpyxl

template_path = r'C:/Users/jjch2/Desktop/템플릿_최종완성본_V73.xlsx'
try:
    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]

    # Write the tags to V73
    ws.cell(row=405, column=2).value = '[[NDT_121_PAUT]]'
    ws.cell(row=450, column=2).value = '[[NDT_RESULT_PAUT]]'

    wb.save(template_path)
    print(f"Tags successfully inserted into {template_path}!")
except Exception as e:
    print(f"Error accessing V73: {e}")
