import openpyxl
from copy import copy

files = [
    r"C:\Users\jjch2\Desktop\누적진도보고서_202708_수정.xlsx",
    r"C:\Users\jjch2\Desktop\템플릿_최종완성본_V70_수정.xlsx"
]

for file_path in files:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        changed = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            cell = ws['G183']
            if cell.value:
                current_font = cell.font
                new_size = 8.0
                
                # We have to create a new Font object
                new_font = copy(current_font)
                new_font.size = new_size
                cell.font = new_font
                
                print(f"[{file_path}] Changed G183 font size in sheet '{sheet_name}' to {new_size}.")
                changed = True
                
        if changed:
            wb.save(file_path)
            print(f"Saved {file_path}")
    except Exception as e:
        print(f"Error processing {file_path}: {e}")
