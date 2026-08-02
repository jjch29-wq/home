import openpyxl
from openpyxl.styles import Border, Side, Alignment

wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')

# 1. Merge D1:D2 for PAUT, PT, MT
for m in ['PAUT', 'PT', 'MT']:
    if m in wb.sheetnames:
        ws = wb[m]
        try:
            ws.merge_cells('D1:D2')
        except ValueError:
            # Already merged or conflict
            pass
        
        # Ensure centered alignment
        ws['D1'].alignment = Alignment(horizontal='center', vertical='center')

# 2. Extend RT to 59 rows
if 'RT' in wb.sheetnames:
    ws_rt = wb['RT']
    hair_side = Side(style='hair')
    thin_side = Side(style='thin')
    
    for r in range(3, 60):
        for c in range(1, 23):
            cell = ws_rt.cell(row=r, column=c)
            left = thin_side if c == 1 else hair_side
            right = thin_side if c == 22 else hair_side
            top = thin_side if r == 3 else hair_side
            bottom = thin_side if r == 59 else hair_side
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            # Add alignment so typed data looks good
            if r > 5:  # For the newly added rows
                cell.alignment = Alignment(horizontal='center', vertical='center')

wb.save(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
print("Merged D1:D2 for PAUT/PT/MT and extended RT to 59 rows.")
