import openpyxl
from openpyxl.styles import Border, Side

wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
ws = wb['RT']

hair_side = Side(style='hair')
thin_side = Side(style='thin')

for r in range(3, ws.max_row + 1):
    for c in range(1, 23):
        left = hair_side
        right = hair_side
        top = hair_side
        bottom = hair_side
        
        if c == 1: left = thin_side
        if c == 22: right = thin_side
        if r == 3: top = thin_side
        if r == ws.max_row: bottom = thin_side
        
        ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

wb.save(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
print("RT inner borders updated to 'hair' style to match other sheets.")
