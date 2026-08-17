import openpyxl

file_path = r"C:\Users\jjch2\Desktop\템플릿_최종완성본_V70.xlsx"
out_path = r"C:\Users\jjch2\Desktop\템플릿_최종완성본_V70_수정.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Clear Z94 if it contains the duplicate sentence
        if ws['Z94'].value and isinstance(ws['Z94'].value, str) and "화성지사" in ws['Z94'].value:
            ws['Z94'].value = None
            
        # Replace headers in B843, B898, B953, and any other row just in case
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "화  성  지  사" in cell.value:
                        cell.value = cell.value.replace("화  성  지  사", "중 앙 지 사")

    wb.save(out_path)
    print(f"Successfully saved to {out_path}")
except Exception as e:
    print(f"Error: {e}")
