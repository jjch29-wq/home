import openpyxl
from openpyxl.styles import Border, Side

wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
ws = wb['RT']

thin_side = Side(style='thin')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

for r in range(3, ws.max_row + 1):
    for c in range(1, 23):
        ws.cell(row=r, column=c).border = thin_border

wb.save(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
print("RT inner borders updated to match other sheets (solid thin).")
