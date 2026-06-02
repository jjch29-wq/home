import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
import threading
import re
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import json

# ============================================================
# Excel Smart Merger v2.8 - Robust Joint/Film/Defect/Rev Quantity Fix
# ============================================================
print("====================================================")
print("  LOADING EXCEL SMART MERGER v2.8 (ROBUST TOTALS)   ")
print("====================================================")

class ExcelMergerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Smart Merger - 지능형 요청서 합치기 v2.8")
        self.root.geometry("850x850")
        self.root.configure(bg="#2c3e50")
        
        self.selected_folder = ""
        self.output_folder = ""
        self.excel_files = []
        
        self.setup_ui()

    def setup_ui(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Main.TFrame", background="#2c3e50")
        
        main_frame = ttk.Frame(self.root, style="Main.TFrame", padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        title_label = tk.Label(main_frame, text="🧠 지능형 엑셀 병합 도구 v2.8", 
                               font=("Malgun Gothic", 18, "bold"), fg="#0be881", bg="#2c3e50")
        title_label.pack(pady=(0, 20))

        # 폴더 선택 (입력 데이터)
        folder_frame = tk.Frame(main_frame, bg="#34495e", padx=10, pady=10)
        folder_frame.pack(fill=tk.X, pady=5)
        self.folder_path_var = tk.StringVar(value="원본 데이터 폴더를 선택해주세요...")
        tk.Label(folder_frame, textvariable=self.folder_path_var, fg="#bdc3c7", bg="#34495e", 
                 font=("Malgun Gothic", 10), anchor="w").pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(folder_frame, text="📄 파일 선택", command=self.select_files, bg="#9b59b6", fg="white", 
                  font=("Malgun Gothic", 10, "bold"), relief=tk.FLAT, padx=15).pack(side=tk.RIGHT, padx=(5, 0))
        tk.Button(folder_frame, text="📁 데이터 폴더 선택", command=self.select_folder, bg="#3498db", fg="white", 
                  font=("Malgun Gothic", 10, "bold"), relief=tk.FLAT, padx=15).pack(side=tk.RIGHT)

        # 저장 폴더 선택 (출력)
        save_folder_frame = tk.Frame(main_frame, bg="#34495e", padx=10, pady=10)
        save_folder_frame.pack(fill=tk.X, pady=5)
        self.output_folder_var = tk.StringVar(value="(선택 안 함) 원본 데이터 폴더에 함께 저장됩니다.")
        tk.Label(save_folder_frame, textvariable=self.output_folder_var, fg="#bdc3c7", bg="#34495e", 
                 font=("Malgun Gothic", 10), anchor="w").pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(save_folder_frame, text="💾 저장 폴더 지정", command=self.select_output_folder, bg="#e67e22", fg="white", 
                  font=("Malgun Gothic", 10, "bold"), relief=tk.FLAT, padx=15).pack(side=tk.RIGHT)

        # 추출 설정
        smart_frame = tk.LabelFrame(main_frame, text=" 추출 설정 (Keywords) ", 
                                    font=("Malgun Gothic", 10, "bold"), fg="#3498db", bg="#2c3e50", padx=10, pady=10)
        smart_frame.pack(fill=tk.X, pady=15)
        
        tk.Label(smart_frame, text="추출할 키워드 (쉼표로 구분):", font=("Malgun Gothic", 9), fg="#ecf0f1", bg="#2c3e50").pack(anchor="w")
        
        self.keyword_var = tk.StringVar(value="No, Joint, Dwg, THK, Result, Date, Report No, Defect Rev")
        keyword_entry = tk.Entry(smart_frame, textvariable=self.keyword_var, font=("Malgun Gothic", 10), bg="#ecf0f1")
        keyword_entry.pack(fill=tk.X, pady=5)
        
        self.only_totals_var = tk.BooleanVar(value=True)
        tk.Checkbutton(smart_frame, text="합계(소계/총합계)만 추출", variable=self.only_totals_var,
                       font=("Malgun Gothic", 9), fg="#ecf0f1", bg="#2c3e50", selectcolor="#2c3e50",
                       activebackground="#2c3e50", activeforeground="#ecf0f1").pack(anchor="w", pady=(2, 2))
                       
        box_label_frame = tk.Frame(smart_frame, bg="#2c3e50")
        box_label_frame.pack(fill=tk.X, pady=(0, 5))
        
        self.export_box_label_var = tk.BooleanVar(value=False)
        tk.Checkbutton(box_label_frame, text="📦 박스라벨 양식으로 자동 생성", variable=self.export_box_label_var,
                       font=("Malgun Gothic", 9, "bold"), fg="#f1c40f", bg="#2c3e50", selectcolor="#2c3e50",
                       activebackground="#2c3e50", activeforeground="#f1c40f").pack(side=tk.LEFT)
                       
        list_frame = tk.Frame(smart_frame, bg="#2c3e50")
        list_frame.pack(fill=tk.X, pady=(5, 5))
        
        self.export_list_var = tk.BooleanVar(value=False)
        tk.Checkbutton(list_frame, text="📋 리스트(목록) 엑셀 자동 생성", variable=self.export_list_var,
                       font=("Malgun Gothic", 9, "bold"), fg="#2ecc71", bg="#2c3e50", selectcolor="#2c3e50",
                       activebackground="#2c3e50", activeforeground="#2ecc71").pack(side=tk.LEFT)
                       
        self.external_list_template_path = ""
        self.btn_browse_list_template = tk.Button(list_frame, text="리스트 양식 찾아보기...", command=self.browse_list_template, bg="#7f8c8d", fg="white", font=("Malgun Gothic", 8), relief=tk.FLAT)
        self.btn_browse_list_template.pack(side=tk.LEFT, padx=(10, 10))
        
        self.list_template_path_var = tk.StringVar(value="(선택 안 함: 내장 기본 양식 사용)")
        tk.Label(list_frame, textvariable=self.list_template_path_var, font=("Malgun Gothic", 8), fg="#bdc3c7", bg="#2c3e50").pack(side=tk.LEFT)
                       
        self.box_label_type_var = tk.StringVar(value="기본 2열 양식")
        self.box_label_combo = ttk.Combobox(box_label_frame, textvariable=self.box_label_type_var, state="readonly", width=20, font=("Malgun Gothic", 9))
        self.box_label_combo['values'] = ("기본 2열 양식", "기본 1열 양식", "외부 엑셀 양식 사용")
        self.box_label_combo.pack(side=tk.LEFT, padx=(10, 5))
        self.box_label_combo.bind("<<ComboboxSelected>>", self.on_box_label_type_selected)
        
        self.external_box_label_path = ""
        self.btn_browse_template = tk.Button(box_label_frame, text="양식 찾아보기...", command=self.browse_external_template, bg="#7f8c8d", fg="white", font=("Malgun Gothic", 8), relief=tk.FLAT)
        self.btn_browse_template.pack(side=tk.LEFT, padx=(0, 10))
        
        self.template_path_var = tk.StringVar(value="")
        tk.Label(box_label_frame, textvariable=self.template_path_var, font=("Malgun Gothic", 8), fg="#bdc3c7", bg="#2c3e50").pack(side=tk.LEFT)
        
        tk.Label(smart_frame, text="💡 v2.8: Joint/Film/Defect/Rev 추출 및 유의어 매칭이 대폭 강화되었습니다.", 
                 font=("Malgun Gothic", 8), fg="#95a5a6", bg="#2c3e50").pack(anchor="w")

        # 시작 버튼 및 상태 (상단 배치)
        self.status_var = tk.StringVar(value="대기 중...")
        tk.Label(main_frame, textvariable=self.status_var, font=("Malgun Gothic", 10, "bold"), fg="#e67e22", bg="#2c3e50").pack(pady=(10, 5))

        self.btn_merge = tk.Button(main_frame, text="🚀 지능형 병합 시작 (v2.8)", command=self.start_merge_thread,
                                   bg="#2ecc71", fg="white", font=("Malgun Gothic", 12, "bold"), relief=tk.FLAT, pady=10)
        self.btn_merge.pack(fill=tk.X, pady=(0, 20))
        self.btn_merge["state"] = tk.DISABLED

        # 실시간 로그창
        log_label = tk.Label(main_frame, text="작업 진행 로그:", font=("Malgun Gothic", 10), fg="#ecf0f1", bg="#2c3e50")
        log_label.pack(anchor="w")
        self.log_text = tk.Text(main_frame, height=20, bg="#1e272e", fg="#0be881", font=("Consolas", 9), padx=10, pady=10)
        self.log_text.pack(fill=tk.BOTH, expand=True, pady=5)

        # 설정 파일 저장 경로
        self.settings_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "smart_merger_settings.json")
        self.load_settings()

    def load_settings(self):
        if os.path.exists(self.settings_file):
            try:
                with open(self.settings_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if "keywords" in data:
                        self.keyword_var.set(data["keywords"])
                    if "only_totals" in data:
                        self.only_totals_var.set(data["only_totals"])
                    if "export_box_label" in data:
                        self.export_box_label_var.set(data["export_box_label"])
                    if "export_list" in data:
                        self.export_list_var.set(data["export_list"])
                    if "box_label_type" in data:
                        self.box_label_type_var.set(data["box_label_type"])
                    if "external_box_label_path" in data:
                        self.external_box_label_path = data["external_box_label_path"]
                        if hasattr(self, 'template_path_var'):
                            self.template_path_var.set(os.path.basename(self.external_box_label_path))
                    if "external_list_template_path" in data:
                        self.external_list_template_path = data["external_list_template_path"]
                        if hasattr(self, 'list_template_path_var') and self.external_list_template_path:
                            self.list_template_path_var.set(os.path.basename(self.external_list_template_path))
                    if "output_folder" in data and data["output_folder"]:
                        self.output_folder = data["output_folder"]
                        self.output_folder_var.set(self.output_folder)
            except Exception as e:
                print("설정 파일 로드 실패:", e)

    def save_settings(self):
        data = {
            "keywords": self.keyword_var.get(),
            "only_totals": self.only_totals_var.get(),
            "export_box_label": self.export_box_label_var.get(),
            "export_list": self.export_list_var.get(),
            "box_label_type": self.box_label_type_var.get(),
            "external_box_label_path": getattr(self, 'external_box_label_path', ""),
            "external_list_template_path": getattr(self, 'external_list_template_path', ""),
            "output_folder": getattr(self, 'output_folder', "")
        }
        try:
            with open(self.settings_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print("설정 파일 저장 실패:", e)

    def browse_external_template(self):
        initial_dir = os.path.dirname(self.external_box_label_path) if self.external_box_label_path and os.path.exists(self.external_box_label_path) else ""
        if not initial_dir:
            initial_dir = os.path.expanduser("~\\Desktop")
            
        filepath = filedialog.askopenfilename(
            title="외부 박스라벨 양식 파일 선택",
            initialdir=initial_dir,
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if filepath:
            self.external_box_label_path = filepath
            self.template_path_var.set(os.path.basename(filepath))
            self.add_log(f"🔗 외부 양식 지정됨: {os.path.basename(filepath)}")
            self.box_label_type_var.set("외부 엑셀 양식 사용")
            self.export_box_label_var.set(True)

    def browse_list_template(self):
        initial_dir = os.path.dirname(self.external_list_template_path) if getattr(self, 'external_list_template_path', "") and os.path.exists(self.external_list_template_path) else ""
        if not initial_dir:
            initial_dir = os.path.expanduser("~\\Desktop")
            
        filepath = filedialog.askopenfilename(
            title="외부 리스트 양식 파일 선택",
            initialdir=initial_dir,
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if filepath:
            self.external_list_template_path = filepath
            self.list_template_path_var.set(os.path.basename(filepath))
            self.add_log(f"🔗 리스트 전용 양식 지정됨: {os.path.basename(filepath)}")
            self.export_list_var.set(True)

    def on_box_label_type_selected(self, event):
        if self.box_label_type_var.get() == "외부 엑셀 양식 사용":
            if not self.external_box_label_path:
                self.browse_external_template()

    def add_log(self, msg):
        timestamp = datetime.now().strftime("[%H:%M:%S] ")
        self.log_text.insert(tk.END, timestamp + msg + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def select_folder(self):
        initial_dir = self.selected_folder if self.selected_folder and os.path.exists(self.selected_folder) else ""
        folder = filedialog.askdirectory(initialdir=initial_dir, title="원본 엑셀 데이터가 있는 폴더 선택")
        if folder:
            self.selected_folder = folder
            self.folder_path_var.set(folder)
            self.scan_files()
            
    def select_output_folder(self):
        initial_dir = self.output_folder if self.output_folder and os.path.exists(self.output_folder) else ""
        folder = filedialog.askdirectory(initialdir=initial_dir, title="최종 결과물을 저장할 폴더 선택")
        if folder:
            self.output_folder = folder
            self.output_folder_var.set(folder)
            self.add_log(f"💾 저장 폴더 지정됨: {folder}")

    def select_files(self):
        initial_dir = self.selected_folder if self.selected_folder and os.path.exists(self.selected_folder) else ""
        files = filedialog.askopenfilenames(
            title="병합할 엑셀 파일들을 선택하세요",
            initialdir=initial_dir,
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
        )
        if files:
            self.selected_folder = os.path.dirname(files[0])
            self.folder_path_var.set(f"개별 파일 선택됨 ({len(files)}개)")
            self.excel_files = [os.path.basename(f) for f in files if "Smart_Merged" not in f and not os.path.basename(f).startswith('~$')]
            
            if self.excel_files:
                self.btn_merge["state"] = tk.NORMAL
                self.status_var.set(f"총 {len(self.excel_files)}개 파일 준비 완료")
                self.add_log(f"개별 파일 선택 완료: {len(self.excel_files)}개")
            else:
                self.add_log("⚠️ 유효한 엑셀 파일이 없습니다.")

    def scan_files(self):
        self.excel_files = [f for f in os.listdir(self.selected_folder) 
                            if (f.endswith('.xlsx') or f.endswith('.xlsm') or f.endswith('.xls')) and not f.startswith('~$') and "Smart_Merged" not in f]
        if self.excel_files:
            self.btn_merge["state"] = tk.NORMAL
            self.status_var.set(f"총 {len(self.excel_files)}개 파일 준비 완료")
            self.add_log(f"파일 검색 완료: {len(self.excel_files)}개")
        else:
            self.add_log("⚠️ 폴더에 엑셀 파일이 없습니다.")

    def start_merge_thread(self):
        self.save_settings()
        self.btn_merge["state"] = tk.DISABLED
        self.log_text.delete("1.0", tk.END)
        threading.Thread(target=self.merge_logic, daemon=True).start()

    def normalize(self, text):
        if pd.isna(text): return ""
        t = str(text).lower()
        t = re.sub(r'[^a-z0-9가-힣]', '', t)
        return t.strip()

    def get_standard_key(self, kw, norm_synonyms):
        if kw in norm_synonyms:
            return kw
        for std_key, syns in norm_synonyms.items():
            if kw in syns:
                return std_key
        return None

    def generate_box_label_openpyxl(self, df_sub, out_path, report_col, joint_col, film_col, defect_col, columns=2):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
        except ImportError:
            self.add_log("⚠️ openpyxl 라이브러리가 없어 박스라벨을 생성할 수 없습니다.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Box Label"
        
        font_bold = Font(name='맑은 고딕', bold=True, size=11)
        font_normal = Font(name='맑은 고딕', size=11)
        align_center = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        
        thin = Side(border_style="thin", color="000000")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 3
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 16
        
        row_idx = 1
        box_num = 1
        
        for i in range(0, len(df_sub), columns):
            page_base = row_idx          # 이 박스 쌍의 첫 번째 행 (상단 상세 라벨)
            side_base = row_idx + 7      # 하단 측면 라벨 시작 행

            items = []
            for j in range(columns):
                r_data = df_sub.iloc[i+j] if i+j < len(df_sub) else None
                items.append((j, j*4, r_data))
            def _write_label_block(base_row, col_offset, row_data, current_box, is_side, joint_val_override=None):
                report_val = str(row_data.get(report_col, '')) if report_col and pd.notna(row_data.get(report_col)) else ''
                joint_count_val = str(int(row_data.get('_joint_count'))) if '_joint_count' in row_data.index and pd.notna(row_data.get('_joint_count')) else (str(row_data.get(joint_col, '')) if joint_col and pd.notna(row_data.get(joint_col)) else '')
                joint_val = joint_val_override if joint_val_override is not None else joint_count_val
                film_val   = str(row_data.get(film_col,   '')) if film_col   and pd.notna(row_data.get(film_col))   else ''
                defect_val = str(row_data.get(defect_col, '')) if defect_col and pd.notna(row_data.get(defect_col)) else ''
                if not str(film_val).strip()   or str(film_val).upper()   == 'NAN': film_val   = '0'
                if not str(defect_val).strip() or str(defect_val).upper() == 'NAN': defect_val = '0'
                if float(film_val)   % 1 == 0 if film_val.replace('.','',1).isdigit()   else False: film_val   = str(int(float(film_val)))
                if float(defect_val) % 1 == 0 if defect_val.replace('.','',1).isdigit() else False: defect_val = str(int(float(defect_val)))
                r = base_row
                c = 1 + col_offset
                current_box_val = row_data.get('_box_num', current_box)
                cell = ws.cell(row=r, column=c, value=f"BOX NO. < {current_box_val} >")
                cell.font = font_bold; cell.alignment = align_center
                ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+2)
                cell = ws.cell(row=r+1, column=c, value="Report No.")
                cell.font = font_bold; cell.alignment = align_center
                cell = ws.cell(row=r+1, column=c+1, value=report_val)
                cell.font = font_normal; cell.alignment = align_center
                ws.merge_cells(start_row=r+1, start_column=c+1, end_row=r+1, end_column=c+2)
                if not is_side:
                    cell = ws.cell(row=r+2, column=c, value="Inspection point.")
                    cell.font = font_bold; cell.alignment = align_center
                    cell = ws.cell(row=r+2, column=c+1, value=joint_val)
                    cell.font = font_normal; cell.alignment = align_center
                    ws.merge_cells(start_row=r+2, start_column=c+1, end_row=r+2, end_column=c+2)
                    cell = ws.cell(row=r+3, column=c, value="필름 매수")
                    cell.font = font_bold; cell.alignment = align_center
                    ws.merge_cells(start_row=r+3, start_column=c, end_row=r+4, end_column=c)
                    cell = ws.cell(row=r+3, column=c+1, value="ORIGINAL")
                    cell.font = font_normal; cell.alignment = align_center
                    cell = ws.cell(row=r+3, column=c+2, value=f"{film_val} 매" if film_val.strip() else "")
                    cell.font = font_normal; cell.alignment = align_center
                    cell = ws.cell(row=r+4, column=c+1, value="REPAIR")
                    cell.font = font_normal; cell.alignment = align_center
                    cell = ws.cell(row=r+4, column=c+2, value=f"{defect_val} 매" if defect_val.strip() else "")
                    cell.font = font_normal; cell.alignment = align_center
                else:
                    # 측면 라벨: 먼저 해당 영역 셀 초기화 (B9 오버플로우 방지)
                    for _r in range(r+2, r+5):
                        for _c in range(c, c+3):
                            ws.cell(row=_r, column=_c).value = None
                    cell = ws.cell(row=r+2, column=c, value="Inspection point.")
                    cell.font = font_bold; cell.alignment = align_center
                    ws.merge_cells(start_row=r+2, start_column=c, end_row=r+4, end_column=c)
                    # joint_val이 비어있을 경우 공백 문자로 채워 오버플로우 방지
                    _jv = joint_val if joint_val.strip() else " "
                    cell = ws.cell(row=r+2, column=c+1, value=_jv)
                    cell.font = Font(name='맑은 고딕', bold=True, size=16)
                    cell.alignment = align_center
                    ws.merge_cells(start_row=r+2, start_column=c+1, end_row=r+4, end_column=c+2)
                try:
                    from openpyxl.cell.rich_text import TextBlock, CellRichText
                    from openpyxl.cell.text import InlineFont
                    sitco_cell = ws.cell(row=r+5, column=c)
                    rf = InlineFont(color='FF0000', rFont='맑은 고딕', b=True, sz=11)
                    bf = InlineFont(color='000000', rFont='맑은 고딕', b=True, sz=11)
                    sitco_cell.value = CellRichText([TextBlock(rf, 'SITCO '), TextBlock(bf, '서울검사(주)')])
                except Exception:
                    sitco_cell = ws.cell(row=r+5, column=c, value="SITCO 서울검사(주)")
                    sitco_cell.font = font_bold
                sitco_cell.alignment = align_center
                ws.merge_cells(start_row=r+5, start_column=c, end_row=r+5, end_column=c+2)
                for row_r in range(r, r+6):
                    ws.row_dimensions[row_r].height = 25
                    for col_c in range(c, c+3):
                        ws.cell(row=row_r, column=col_c).border = border_all
                return joint_val  # detail 패스에서 저장용

            saved_joint_vals = {}
            for j, col_offset, row_data in items:
                if row_data is None: continue
                jv = _write_label_block(page_base, col_offset, row_data, box_num + j, is_side=False)
                saved_joint_vals[j] = jv
            for j, col_offset, row_data in items:
                if row_data is None: continue
                _write_label_block(side_base, col_offset, row_data, box_num + j, is_side=True,
                                   joint_val_override=saved_joint_vals.get(j))

            row_idx += 14
            box_num += columns

        ws.print_area = f"A1:G{row_idx-1}"
        wb.save(out_path)
        self.add_log(f"📦 박스라벨(새 양식) 생성 완료: {os.path.basename(out_path)}")

    def generate_box_label(self, df_merged, timestamp):
        if '_is_total' in df_merged.columns:
            df_sub = df_merged[df_merged['_is_total'] == True].copy()
        else:
            mask = df_merged.astype(str).apply(lambda row: row.str.contains("Sub-Total|소계", case=False).any(), axis=1)
            df_sub = df_merged[mask].copy()

        df_sub = df_sub[~df_sub.astype(str).apply(lambda row: row.str.contains("Grand Total|총합계", case=False).any(), axis=1)]
        
        def get_col(synonyms, default=None):
            return next((c for c in df_sub.columns if str(c).lower().strip() in synonyms or any(s in str(c).lower() for s in synonyms)), default)

        report_col = get_col(['report', '성적서'])
        joint_col = get_col(['joint', '조인트'])
        film_col = get_col(['film', '필름'])
        defect_col = get_col(['defect', '결함', 'repair'])
        
        # 성적서 번호(Report No) 기준으로 알파벳/숫자 오름차순 정렬
        # 이를 통해 0022 -> 0022R1 순서가 자연스럽게 이어지도록 함
        if report_col:
            df_sub = df_sub.sort_values(by=report_col, na_position='last').reset_index(drop=True)
            
        # R1 보고서의 경우 REPAIR(결함) 수량 강제로 0매 처리 (단, R1-1, R1-2 등은 제외)
        if report_col and defect_col:
            for idx, row in df_sub.iterrows():
                if pd.notna(row[report_col]):
                    r_str = str(row[report_col]).upper()
                    if "R1" in r_str and not re.search(r'R1-\d+', r_str):
                        df_sub.at[idx, defect_col] = 0
                        

        # 100매 초과 시 박스 분할 로직
        if film_col:
            expanded_rows = []
            box_counter = 1
            for _, row in df_sub.iterrows():
                try:
                    f_val = str(row[film_col]).replace(',', '').strip()
                    film_count = float(f_val) if f_val else 0
                except:
                    film_count = 0
                    
                if film_count > 100:
                    import math
                    num_splits = math.ceil(film_count / 100)
                    for i in range(num_splits):
                        new_row = row.copy()
                        # 보고서 번호는 동일하게 유지
                        
                        # 분할된 나머지 박스들은 결함(REPAIR) 수량을 0으로 처리
                        if i > 0 and defect_col:
                            new_row[defect_col] = 0
                        
                        if i == num_splits - 1:
                            new_row[film_col] = int(film_count - (i * 100))
                        else:
                            new_row[film_col] = 100
                            
                        # 박스 번호에 -1, -2 붙임 (예: 2-1, 2-2)
                        new_row['_box_num'] = f"{box_counter}-{i+1}"
                        expanded_rows.append(new_row)
                else:
                    new_row = row.copy()
                    new_row['_box_num'] = box_counter
                    expanded_rows.append(new_row)
                
                box_counter += 1
            df_sub = pd.DataFrame(expanded_rows)
        else:
            df_sub['_box_num'] = range(1, len(df_sub) + 1)
            
        save_dir = self.output_folder if hasattr(self, 'output_folder') and self.output_folder else self.selected_folder
        
        # 리스트 생성 (체크된 경우)
        if hasattr(self, 'export_list_var') and self.export_list_var.get():
            self.add_log("📋 리스트(목록) 생성을 시작합니다...")
            self.generate_list_excel(df_sub, save_dir, timestamp, report_col, joint_col, film_col, defect_col)

        if not self.export_box_label_var.get():
            return
        
        box_type = self.box_label_type_var.get()
        
        if box_type == "기본 1열 양식":
            out_path = os.path.join(save_dir, f"Final_BoxLabel_{timestamp}.xlsx")
            self.generate_box_label_openpyxl(df_sub, out_path, report_col, joint_col, film_col, defect_col, columns=1)
            return
            
        elif box_type == "기본 2열 양식":
            out_path = os.path.join(save_dir, f"Final_BoxLabel_{timestamp}.xlsx")
            self.generate_box_label_openpyxl(df_sub, out_path, report_col, joint_col, film_col, defect_col, columns=2)
            return
            
        # 외부 엑셀 양식 처리 로직
        if not self.external_box_label_path or not os.path.exists(self.external_box_label_path):
            self.add_log(f"⚠️ 선택된 외부 템플릿을 찾을 수 없어 기본 양식으로 생성합니다.")
            out_path = os.path.join(save_dir, f"Final_BoxLabel_{timestamp}.xlsx")
            self.generate_box_label_openpyxl(df_sub, out_path, report_col, joint_col, film_col, defect_col, columns=2)
            return

        template_path = os.path.normpath(os.path.abspath(self.external_box_label_path))

        try:
            import win32com.client as win32
        except ImportError:
            out_path = os.path.join(save_dir, f"Final_BoxLabel_{timestamp}.xlsx")
            self.add_log("⚠️ pywin32(엑셀 제어) 패키지가 없어 기본 양식으로 생성합니다.")
            self.generate_box_label_openpyxl(df_sub, out_path, report_col, joint_col, film_col, defect_col, columns=2)
            return

        self.add_log(f"📦 원본 템플릿({os.path.basename(template_path)})을 사용하여 박스라벨 생성을 시작합니다...")
        
        excel = None
        wb = None
        out_name = f"Final_BoxLabel_원본양식_{timestamp}.xlsx"
        out_path = os.path.normpath(os.path.abspath(os.path.join(save_dir, out_name)))
        
        try:
            excel = win32.DispatchEx('Excel.Application')
            try: excel.Visible = False
            except: pass
            try: excel.DisplayAlerts = False
            except: pass
            
            wb = excel.Workbooks.Open(template_path)
            try: ws = wb.Sheets('2021')
            except: ws = wb.Sheets(1)
            
            template_range = ws.Range("A1:G25")
            ws.Range("A26:G5000").Clear()
            
            boxes = [df_sub.iloc[i] for i in range(len(df_sub))]
            box_num = 1
            
            for p in range(0, len(boxes), 4):
                page_boxes = boxes[p:p+4]
                page_start_row = 1 + (p // 4) * 25
                
                if p > 0:
                    template_range.Copy()
                    ws.Cells(page_start_row, 1).PasteSpecial(Paste=-4104) # xlPasteAll
                    
                    # 첫 페이지의 행 높이 복사
                    for i in range(25):
                        ws.Rows(page_start_row + i).RowHeight = ws.Rows(i + 1).RowHeight
                        
                for idx, row_data in enumerate(page_boxes):
                    if idx < 2:
                        base_row = page_start_row
                        col_offset = 0 if idx == 0 else 4
                    else:
                        base_row = page_start_row + 13
                        col_offset = 0 if idx == 2 else 4
                        
                    current_box = box_num + idx
                    
                    report_val = str(row_data.get(report_col, '')) if report_col and pd.notna(row_data.get(report_col)) else ''
                    joint_val = str(int(row_data.get('_joint_count'))) if '_joint_count' in row_data.index and pd.notna(row_data.get('_joint_count')) else (str(row_data.get(joint_col, '')) if joint_col and pd.notna(row_data.get(joint_col)) else '')
                    self.add_log(f"   🔍 [DEBUG] BOX {current_box}: _joint_count={'있음:'+str(row_data.get('_joint_count')) if '_joint_count' in row_data.index else '없음'}, joint_col={joint_col}, joint_val={joint_val}")
                    film_val = str(row_data.get(film_col, '')) if film_col and pd.notna(row_data.get(film_col)) else ''
                    defect_val = str(row_data.get(defect_col, '')) if defect_col and pd.notna(row_data.get(defect_col)) else ''
                    
                    if not str(film_val).strip() or str(film_val).upper() == 'NAN': film_val = '0'
                    if not str(defect_val).strip() or str(defect_val).upper() == 'NAN': defect_val = '0'
                    
                    if float(film_val) % 1 == 0 if film_val.replace('.','',1).isdigit() else False: film_val = str(int(float(film_val)))
                    if float(defect_val) % 1 == 0 if defect_val.replace('.','',1).isdigit() else False: defect_val = str(int(float(defect_val)))
                    
                    for repeat in range(2):
                        r = base_row + repeat * 7
                        current_box_val = row_data.get('_box_num', current_box)
                        ws.Cells(r, 1 + col_offset).Value = f"BOX NO. < {current_box_val} >"
                        ws.Cells(r+1, 2 + col_offset).Value = report_val
                        
                        if repeat == 0:
                            ws.Cells(r+2, 2 + col_offset).Value = joint_val
                            ws.Cells(r+3, 3 + col_offset).Value = film_val
                            ws.Cells(r+4, 3 + col_offset).Value = defect_val
                        else:
                            # 하단 측면 라벨 — 기존 템플릿에 남아있을 수 있는 값/수식 완벽히 지우기
                            ws.Range(ws.Cells(r+2, 2 + col_offset), ws.Cells(r+4, 3 + col_offset)).ClearContents()
                            # 좌상단 셀에 값 입력
                            ws.Cells(r+2, 2 + col_offset).Value = str(joint_val)
                        
                if len(page_boxes) == 1:
                    ws.Range(ws.Cells(page_start_row, 5), ws.Cells(page_start_row+12, 7)).Clear()
                    ws.Range(ws.Cells(page_start_row+13, 1), ws.Cells(page_start_row+24, 7)).Clear()
                elif len(page_boxes) == 2:
                    ws.Range(ws.Cells(page_start_row+13, 1), ws.Cells(page_start_row+24, 7)).Clear()
                elif len(page_boxes) == 3:
                    ws.Range(ws.Cells(page_start_row+13, 5), ws.Cells(page_start_row+24, 7)).Clear()
                    
                box_num += len(page_boxes)
                
            last_row = 25 + ((max(0, len(boxes) - 1)) // 4) * 25
            ws.PageSetup.PrintArea = f"A1:G{last_row}"
                
            wb.SaveAs(out_path, FileFormat=51) # xlOpenXMLWorkbook (.xlsx)
            wb.Close(SaveChanges=False)
            excel.Quit()
            
            self.add_log(f"✨ 원본 양식 기반 박스라벨 생성 완료: {out_name}")
            
        except Exception as e:
            self.add_log(f"❌ 원본 박스라벨 엑셀 제어 중 오류 발생: {e}")
            if wb:
                try: wb.Close(SaveChanges=False)
                except: pass
            if excel:
                try: excel.Quit()
                except: pass

    def merge_logic(self):
        try:
            all_data = []
            raw_kws = [k.strip() for k in self.keyword_var.get().split(',') if k.strip()]
            norm_keywords = [self.normalize(k) for k in raw_kws]
            
            if not norm_keywords:
                messagebox.showwarning("알림", "검색할 키워드를 입력해주세요.")
                return

            self.add_log("🚀 v2.8 지능형 병합 엔진 기동 중...")

            # 1. 유의어(Synonyms) 사전 정의
            SYNONYMS = {
                "no": ["no", "no.", "seq", "순번", "연번", "일련번호", "번호", "num", "idx", "호"],
                "dwg": ["dwg", "dwgno", "dwg.no", "도면", "도면번호", "도면명", "drawing", "drawingno", "drawing.no", "iso"],
                "joint": ["joint", "jointno", "joint.no", "조인트", "jnt", "jntno", "point", "포인트"],
                "size": ["size", "규격", "구경", "사이즈", "dia", "nps", "pipe size", "파이프 규격"],
                "thk": ["thk", "thickness", "두께", "t", "thk.", "thick"],
                "result": ["result", "결과", "판정", "결과판정", "decision", "판정결과"],
                "date": ["date", "날짜", "검사일", "검사일자", "일자", "dateofexam", "요청일", "요청일자"],
                "reportno": ["reportno", "report.no", "report_no", "성적서번호", "성적서", "보고서번호", "보고서"],
                "identificationno": ["identificationno", "idno", "id_no", "관리번호", "식별번호", "id"],
                "film": ["film", "filmno", "film.no", "필름", "필름번호", "매수", "수량", "qty", "quantity", "filmqty", "filmquantity", "nooffilm", "numberoffilm", "jointqty", "jointquantity", "joint수량", "조인트수량"],
                "defect": ["defect", "reject", "불합격", "결함", "defectqty", "rejectqty", "불합격수량", "결함수량", "defect rev", "defectrev", "defect_rev"]
            }
            
            # 정규화된 유의어 맵 생성
            NORM_SYNONYMS = {k: [self.normalize(s) for s in v] for k, v in SYNONYMS.items()}
            
            # 제목 줄 찾기 점수 계산용 핵심 키워드 리스트
            header_kws = ['no', 'joint', 'dwg', 'size', 'thk', 'result', 'date']

            for file in self.excel_files:
                self.add_log(f"--- 분석: {file} ---")
                file_path = os.path.join(self.selected_folder, file)
                
                try:
                    xls = pd.ExcelFile(file_path)
                    
                    for sheet_name in xls.sheet_names:
                        raw_df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                        
                        # 1. 전역 메타데이터 (Report No) 추출
                        meta_info = {}
                        for r_idx, row in raw_df.iterrows():
                            if r_idx > 50: break
                            row_vals = row.values
                            for c_idx, val in enumerate(row_vals):
                                if pd.isna(val): continue
                                s_val = str(val)
                                s_upper = s_val.upper()
                                if ("REPORT" in s_upper and "NO" in s_upper) or ("성적서" in s_val and "번호" in s_val):
                                    extracted_val = ""
                                    if ":" in s_val:
                                        extracted_val = s_val.split(":", 1)[1].strip()
                                    else:
                                        # 셀 내에 숫자나 전형적인 번호 형식이 없으면 단지 라벨로 판단
                                        if not re.search(r'\d', s_val):
                                            extracted_val = ""
                                        else:
                                            tmp = re.sub(r'(?i)(report\s*no\.?|성적서\s*번호)', '', s_val).strip()
                                            if tmp: extracted_val = tmp
                                    
                                    if not extracted_val or extracted_val.upper() == "NAN":
                                        for offset in range(1, 4):
                                            if c_idx + offset < len(row_vals):
                                                v = str(row_vals[c_idx + offset]).strip()
                                                if v and v.upper() != "NAN" and len(v) < 30:
                                                    extracted_val = v
                                                    break
                                    
                                    if extracted_val and extracted_val.upper() != "NAN":
                                        clean_val = re.split(r'(?i)\n|/|\s{2,}|date|일자|page|페이지|\(|\[|<', extracted_val)[0].strip()
                                        clean_val = re.split(r'(?i)\s+[가-힣]{2,}|\s+(?:rev|sheet|insp|note|remark)', clean_val)[0].strip()
                                        if ":" in clean_val:
                                            tmp = clean_val.split(":")[0].strip()
                                            clean_val = re.sub(r'\s+[A-Za-z가-힣]+$', '', tmp).strip()
                                        clean_val = re.split(r'(?i)\s+(?:IP|LF|CR|PO|UC|BT|NF|INCOMPLETE|LACK|DEFECT|결함)\b', clean_val)[0].strip()
                                        clean_val = re.sub(r'^[:\-]+|[:\-]+$', '', clean_val).strip()
                                        
                                        if clean_val:
                                            if re.search(r'\d', clean_val) or clean_val.upper() in ['N/A', '-', 'TBD', 'NA']:
                                                meta_info["Report No"] = clean_val
                                                self.add_log(f"   📌 메타데이터 정밀 추출: Report No -> {clean_val}")
                                                break
                                            else:
                                                extracted_val = ""
                            if "Report No" in meta_info: break
 
                        if "Report No" not in meta_info:
                            base_name = os.path.splitext(file)[0]
                            meta_info["Report No"] = base_name
                            self.add_log(f"   📌 메타데이터 (파일명 기준): Report No -> {base_name}")

                        # 2. 제목 줄 찾기 (유의어 기반 고도화 점수제)
                        best_row = 0
                        max_score = 0
                        for idx, row in raw_df.iterrows():
                            if idx > 60: break
                            row_content = "".join([str(v) for v in row.values if pd.notna(v)])
                            norm_content = self.normalize(row_content)
                            score = 0
                            for kw in header_kws:
                                norm_kw = self.normalize(kw)
                                if norm_kw in norm_content or any(syn in norm_content for syn in NORM_SYNONYMS.get(norm_kw, [])):
                                    score += 1
                            if score > max_score:
                                max_score = score
                                best_row = idx
                        
                        # 최고점 행 이후 더 완전한 헤더 행 탐색 (Row 0이 불완전하고 Row 2가 완전한 케이스 대응)
                        final_header_row = best_row
                        best_row_nonnull = sum(1 for v in raw_df.iloc[best_row] if pd.notna(v) and str(v).strip() not in ['', 'nan'])
                        for look_ahead in range(1, 4):
                            check_idx = best_row + look_ahead
                            if check_idx >= len(raw_df):
                                break
                            row_content = "".join([str(v) for v in raw_df.iloc[check_idx].values if pd.notna(v)])
                            norm_content = self.normalize(row_content)
                            row_score = sum(1 for kw in header_kws if self.normalize(kw) in norm_content or any(syn in norm_content for syn in NORM_SYNONYMS.get(self.normalize(kw), [])))
                            row_nonnull = sum(1 for v in raw_df.iloc[check_idx] if pd.notna(v) and str(v).strip() not in ['', 'nan'])
                            if row_score >= max_score and row_nonnull > best_row_nonnull:
                                final_header_row = check_idx
                                best_row_nonnull = row_nonnull
                        
                        # 유효한 테이블 감지 조건 (점수 3점 이상)
                        if max_score >= 3:
                            df = pd.read_excel(file_path, sheet_name=sheet_name, header=final_header_row)
                            
                            # 컬럼 정리 및 중복 방어
                            df.columns = [str(c).strip() if not str(c).startswith('Unnamed') else f"_col{i}" for i, c in enumerate(df.columns)]
                            df = df.loc[:, ~df.columns.duplicated()]
                            
                            # 최종 컬럼 유의어 기반 정밀 매칭 (1키워드 당 1컬럼 매칭)
                            final_cols = []
                            norm_col_map = {col: self.normalize(col) for col in df.columns}
                            keyword_to_col = {}
                            
                            for raw_kw in raw_kws:
                                kw = self.normalize(raw_kw)
                                std_key = self.get_standard_key(kw, NORM_SYNONYMS)
                                
                                # 1. 정확 매칭
                                match = next((orig for orig, norm in norm_col_map.items() if norm == kw and orig not in final_cols), None)
                                if match:
                                    final_cols.append(match)
                                    keyword_to_col[kw] = match
                                    continue
                                
                                # 2. 유의어 매칭
                                if std_key:
                                    syns = NORM_SYNONYMS.get(std_key, [])
                                    match = next((orig for orig, norm in norm_col_map.items() if norm in syns and orig not in final_cols), None)
                                    if match:
                                        final_cols.append(match)
                                        keyword_to_col[kw] = match
                                        continue
                                    
                                # 3. 유사(부분) 매칭 (너무 짧은 단어 오매칭 방지 및 synonym 길이 제한)
                                if len(kw) >= 3:
                                    syns = NORM_SYNONYMS.get(std_key, []) if std_key else []
                                    match = next((orig for orig, norm in norm_col_map.items() if (kw in norm or any(syn in norm for syn in syns if len(syn) >= 3)) and orig not in final_cols), None)
                                    if match:
                                        final_cols.append(match)
                                        keyword_to_col[kw] = match
                                        continue

                            # 핵심 Joint 컬럼이 없으면 본문 데이터가 없는 갑지/요약 시트로 판정하여 스킵
                            joint_matched = any(self.get_standard_key(k, NORM_SYNONYMS) == "joint" for k in keyword_to_col.keys())
                            if not joint_matched:
                                self.add_log(f"   ⚠️ 시트 스킵: '{sheet_name}' (Joint 컬럼 유실)")
                                continue
                                
                            # 1. 쓰레기 데이터 정리 (ffill 적용 전, 원본 컬럼명 기준)
                            raw_no_col = next((c for c, norm in norm_col_map.items() if norm == "no" or norm in NORM_SYNONYMS.get("no", [])), None)
                            raw_joint_col = next((c for c, norm in norm_col_map.items() if norm == "joint" or norm in NORM_SYNONYMS.get("joint", [])), None)
                            
                            filter_col_raw = raw_no_col if raw_no_col else raw_joint_col
                            
                            if filter_col_raw:
                                df = df.dropna(subset=[filter_col_raw])
                                df = df[df[filter_col_raw].astype(str).str.strip() != ""]
                                
                                if filter_col_raw == raw_no_col:
                                    # 순번 필터: 숫자를 포함해야 하되, 하단 정보란이나 불필요 서명 배제
                                    df = df[df[raw_no_col].astype(str).str.contains(r'\d', regex=True, na=False)]
                                    exclude_terms = r'(?i)page|페이지|total|합계|sub-total|소계|grand|seoul|inspection|testing|report|project|client|customer|date'
                                    df = df[~df[raw_no_col].astype(str).str.contains(exclude_terms, regex=True, na=False)]
                                else:
                                    # 조인트 필터: 하단 서명란, 빈 줄 꼬리표 및 중복 헤더 단어(Joint) 제거
                                    exclude_terms = r'(?i)page|페이지|total|합계|sub-total|소계|grand|seoul|inspection|testing|report|project|client|customer|date|joint|조인트'
                                    df = df[~df[filter_col_raw].astype(str).str.contains(exclude_terms, regex=True, na=False)]
                                    
                            self.add_log(f"   🎯 테이블 감지됨: '{sheet_name}' 시트, {best_row+1}행 -> {len(df)}행 추출 완료")

                            # 2. 매칭 컬럼 필터링 및 표준 키워드로 컬럼명 변경 (병합 시 완벽 정렬 보장!)
                            df = df[final_cols]
                            rename_map = {keyword_to_col[kw]: raw_kw for raw_kw in raw_kws if (kw := self.normalize(raw_kw)) in keyword_to_col}
                            df.rename(columns=rename_map, inplace=True)
                            
                            # 3. 유의어 매칭을 사용해 실제 리네임된 도면, 조인트 컬럼 찾기
                            dwg_col_real = next((c for c in df.columns if self.normalize(c) == "dwg" or any(self.normalize(c) == syn for syn in NORM_SYNONYMS.get("dwg", []))), None)
                            joint_col_real = next((c for c in df.columns if self.normalize(c) == "joint" or any(self.normalize(c) == syn for syn in NORM_SYNONYMS.get("joint", []))), None)
                            
                            # 4. 세로 병합 셀(Merged Cells) 복원을 위한 Forward Fill 적용
                            if dwg_col_real:
                                df[dwg_col_real] = df[dwg_col_real].ffill()
                            if joint_col_real:
                                df[joint_col_real] = df[joint_col_real].ffill()
                            
                            # 메타데이터 주입
                            for m_key, m_val in meta_info.items():
                                df[m_key] = m_val
                                
                            all_data.append(df)
                            
                except Exception as fe:
                    self.add_log(f"   ❌ 시트 분석 중 오류 발생: {str(fe)}")

            if not all_data:
                self.add_log("❌ 병합할 데이터가 없습니다.")
                return

            self.add_log("📊 전체 데이터 병합 및 단일 매칭 필터링 중...")
            combined_df = pd.concat(all_data, ignore_index=True, sort=False)
            combined_df = combined_df.loc[:, ~combined_df.columns.duplicated()]
            
            # 주의: drop_duplicates 제거 — 같은 Joint 번호가 다른 Dwg/날짜에 걸쳐 여러 번 나올 수 있어 중복 제거 금지
            # combined_df.drop_duplicates(inplace=True)
            
            # No. of Film 수량 원본 파일별 소계 및 전체 총합산 기능
            film_col = next((c for c in combined_df.columns if self.normalize(c) == "film" or any(self.normalize(c) == syn for syn in NORM_SYNONYMS.get("film", []))), None)
            joint_col = next((c for c in combined_df.columns if self.normalize(c) == "joint" or any(self.normalize(c) == syn for syn in NORM_SYNONYMS.get("joint", []))), None)
            defect_col = next((c for c in combined_df.columns if self.normalize(c) == "defect" or any(self.normalize(c) == syn for syn in NORM_SYNONYMS.get("defect", []))), None)
            rev_col = next((c for c in combined_df.columns if self.normalize(c) == "rev" or any(self.normalize(c) == syn for syn in NORM_SYNONYMS.get("rev", []))), None)
            
            if len(combined_df) > 0:
                no_col_name = next((c for c in combined_df.columns if self.normalize(c) == "no"), None)
                
                report_col = next((c for c in combined_df.columns if self.normalize(c) == "reportno" or any(syn in self.normalize(c) for syn in NORM_SYNONYMS.get("reportno", []))), None)
                # label_col: Sub-Total 텍스트를 넣을 컬럼 — 항상 joint_col, 없으면 report_col
                label_col = joint_col if joint_col else report_col
                
                if film_col:
                    combined_df['_temp_numeric_film'] = pd.to_numeric(combined_df[film_col].astype(str).str.extract(r'(\d+\.?\d*)')[0], errors='coerce')
                if defect_col:
                    combined_df['_temp_numeric_defect'] = pd.to_numeric(combined_df[defect_col].astype(str).str.extract(r'(\d+\.?\d*)')[0], errors='coerce')
                if rev_col:
                    combined_df['_temp_numeric_rev'] = pd.to_numeric(combined_df[rev_col].astype(str).str.extract(r'(\d+\.?\d*)')[0], errors='coerce')
                
                new_dfs = []
                grand_total = 0
                grand_joint_total = 0
                grand_defect_total = 0
                grand_rev_total = 0
                
                if report_col:
                    for rep_no, group in combined_df.groupby(report_col, sort=False):
                        new_dfs.append(group)
                        
                        sub_total = group['_temp_numeric_film'].sum() if film_col else 0
                        grand_total += sub_total
                        
                        sub_joint = 0
                        if joint_col:
                            valid_joints = group[joint_col].dropna().astype(str).str.strip()
                            valid_joints = valid_joints[valid_joints != ""]
                            sub_joint = valid_joints.count()
                        grand_joint_total += sub_joint
                        
                        # Defect calculation
                        sub_defect = 0
                        if defect_col:
                            num_sum = group['_temp_numeric_defect'].sum()
                            if num_sum > 0:
                                sub_defect = num_sum
                            else:
                                valid_vals = group[defect_col].dropna().astype(str).str.strip().str.lower()
                                sub_defect = valid_vals[valid_vals.str.contains('reject|fail|ng|불합격|결함', na=False)].count()
                        grand_defect_total += sub_defect

                        # Rev calculation
                        sub_rev = 0
                        if rev_col:
                            num_sum = group['_temp_numeric_rev'].sum()
                            if num_sum > 0:
                                sub_rev = num_sum
                            else:
                                valid_vals = group[rev_col].dropna().astype(str).str.strip().str.lower()
                                sub_rev = valid_vals[valid_vals.str.contains(r'^r|repair|보수|수리|개정', na=False)].count()
                        grand_rev_total += sub_rev
                        
                        sub_row = {col: "" for col in combined_df.columns}
                        sub_row['_is_total'] = True
                        # Sub-Total: Report No 컬럼에 성적서 번호만 표기, Joint에 집계 숫자
                        if report_col:
                            sub_row[report_col] = f"{rep_no}"
                        if joint_col:
                            sub_row[joint_col] = int(sub_joint)
                        sub_row['_joint_count'] = int(sub_joint)
                        if film_col:
                            sub_row[film_col] = int(sub_total) if isinstance(sub_total, float) and sub_total % 1 == 0 else sub_total
                        if defect_col:
                            if sub_defect > 0:
                                sub_row[defect_col] = int(sub_defect) if isinstance(sub_defect, float) and sub_defect % 1 == 0 else sub_defect
                            else:
                                sub_row[defect_col] = ""
                        if rev_col:
                            if sub_rev > 0:
                                sub_row[rev_col] = int(sub_rev) if isinstance(sub_rev, float) and sub_rev % 1 == 0 else sub_rev
                            else:
                                sub_row[rev_col] = ""
                        new_dfs.append(pd.DataFrame([sub_row]))
                    
                    combined_df = pd.concat(new_dfs, ignore_index=True)
                else:
                    grand_total = combined_df['_temp_numeric_film'].sum() if film_col else 0
                    if joint_col:
                        valid_joints = combined_df[joint_col].dropna().astype(str).str.strip()
                        valid_joints = valid_joints[valid_joints != ""]
                        grand_joint_total = valid_joints.count()
                    else:
                        grand_joint_total = 0
                        
                    if defect_col:
                        num_sum = combined_df['_temp_numeric_defect'].sum()
                        if num_sum > 0:
                            grand_defect_total = num_sum
                        else:
                            valid_vals = combined_df[defect_col].dropna().astype(str).str.strip().str.lower()
                            grand_defect_total = valid_vals[valid_vals.str.contains('reject|fail|ng|불합격|결함', na=False)].count()
                    else:
                        grand_defect_total = 0
                        
                    if rev_col:
                        num_sum = combined_df['_temp_numeric_rev'].sum()
                        if num_sum > 0:
                            grand_rev_total = num_sum
                        else:
                            valid_vals = combined_df[rev_col].dropna().astype(str).str.strip().str.lower()
                            grand_rev_total = valid_vals[valid_vals.str.contains(r'^r|repair|보수|수리|개정', na=False)].count()
                    else:
                        grand_rev_total = 0

                total_row = {col: "" for col in combined_df.columns}
                total_row['_is_total'] = True
                # Grand Total: Report No 컬럼에 'Grand Total', Joint에 총합계 숫자
                if report_col:
                    total_row[report_col] = "Grand Total"
                if joint_col:
                    total_row[joint_col] = int(grand_joint_total)
                if film_col:
                    total_row[film_col] = int(grand_total) if isinstance(grand_total, float) and grand_total % 1 == 0 else grand_total
                if defect_col:
                    if grand_defect_total > 0:
                        total_row[defect_col] = int(grand_defect_total) if isinstance(grand_defect_total, float) and grand_defect_total % 1 == 0 else grand_defect_total
                    else:
                        total_row[defect_col] = ""
                if rev_col:
                    if grand_rev_total > 0:
                        total_row[rev_col] = int(grand_rev_total) if isinstance(grand_rev_total, float) and grand_rev_total % 1 == 0 else grand_rev_total
                    else:
                        total_row[rev_col] = ""
                combined_df = pd.concat([combined_df, pd.DataFrame([total_row])], ignore_index=True)
                
                log_msg = "   ➕ 소계 및 총합계 추가 완료"
                if film_col or joint_col or defect_col or rev_col:
                    log_msg += " (수량 합산 포함)"
                self.add_log(log_msg)
                
                if film_col and '_temp_numeric_film' in combined_df.columns:
                    combined_df.drop(columns=['_temp_numeric_film'], inplace=True)
                if defect_col and '_temp_numeric_defect' in combined_df.columns:
                    combined_df.drop(columns=['_temp_numeric_defect'], inplace=True)
                if rev_col and '_temp_numeric_rev' in combined_df.columns:
                    combined_df.drop(columns=['_temp_numeric_rev'], inplace=True)
                
                # Defect & Rev 컬럼에서 1 미만(0 또는 0.0) 값은 빈칸으로 처리하여 깔끔하게 표시
                for col in [defect_col, rev_col]:
                    if col and col in combined_df.columns:
                        def _clean_zero_val(val):
                            if pd.isna(val) or str(val).strip() == "":
                                return ""
                            try:
                                num = pd.to_numeric(val)
                                if num >= 1:
                                    return int(num) if num % 1 == 0 else num
                                else:
                                    return ""
                            except:
                                return val
                        combined_df[col] = combined_df[col].apply(_clean_zero_val)
                
            timestamp = datetime.now().strftime('%H%M%S')
            
            if self.export_box_label_var.get():
                self.generate_box_label(combined_df, timestamp)
            
            if self.only_totals_var.get() and '_is_total' in combined_df.columns:
                combined_df = combined_df[combined_df['_is_total'] == True]
                
            if '_is_total' in combined_df.columns:
                combined_df.drop(columns=['_is_total'], inplace=True)

            save_dir = self.output_folder if hasattr(self, 'output_folder') and self.output_folder else self.selected_folder
            out_name = f"Smart_Merged_Report_{timestamp}.xlsx"
            out_path = os.path.join(save_dir, out_name)
            combined_df.to_excel(out_path, index=False)
            
            self.add_log(f"✨ 완료: {out_name}")
            self.status_var.set(f"✅ 저장 완료: {out_name}")
            messagebox.showinfo("성공", f"지능형 병합이 성공적으로 완료되었습니다!\n파일명: {out_name}")
            
        except Exception as e:
            self.add_log(f"❌ 치명적 오류: {str(e)}")
            messagebox.showerror("오류", f"프로세스 오류: {e}")
        finally:
            self.btn_merge["state"] = tk.NORMAL

    def generate_list_excel(self, df_sub, save_dir, timestamp, report_col, joint_col, film_col, defect_col):
        list_data = []
        for idx, row in df_sub.iterrows():
            box_no = str(row.get('_box_num', ''))
            report_no = str(row[report_col]) if report_col and pd.notna(row[report_col]) else ''
            joint = str(row[joint_col]) if joint_col and pd.notna(row[joint_col]) else ''
            film = str(row[film_col]) if film_col and pd.notna(row[film_col]) else ''
            defect = str(row[defect_col]) if defect_col and pd.notna(row[defect_col]) else '0'
            
            # Format numbers to remove trailing .0
            if film.replace('.','',1).isdigit() and float(film) % 1 == 0: film = str(int(float(film)))
            if defect.replace('.','',1).isdigit() and float(defect) % 1 == 0: defect = str(int(float(defect)))
            if joint.replace('.','',1).isdigit() and float(joint) % 1 == 0: joint = str(int(float(joint)))
            
            list_data.append({
                "No.": idx + 1,
                "BOX NO.": box_no,
                "Report No.": report_no,
                "Joint": joint,
                "Film": film,
                "Repair": defect,
                "Remark": ""
            })
            
        df_list = pd.DataFrame(list_data)
        out_name = f"Final_List_{timestamp}.xlsx"
        
        # 외부 템플릿 사용
        if hasattr(self, 'external_list_template_path') and self.external_list_template_path and os.path.exists(self.external_list_template_path):
            template_path = os.path.normpath(os.path.abspath(self.external_list_template_path))
            out_path = os.path.normpath(os.path.abspath(os.path.join(save_dir, f"Final_List_외부양식_{timestamp}.xlsx")))
            
            try:
                import win32com.client as win32
                excel = win32.DispatchEx('Excel.Application')
                excel.Visible = False
                excel.DisplayAlerts = False
                wb = excel.Workbooks.Open(template_path)
                ws = wb.Sheets(1)
                
                for i, item in enumerate(list_data):
                    page = i // 40
                    pos = i % 40
                    
                    page_start_row = 1 + page * 23
                    
                    # 새로운 페이지 생성
                    if page > 0 and pos == 0:
                        ws.Range("A1:L23").Copy()
                        ws.Range(f"A{page_start_row}").PasteSpecial(-4104) # xlPasteAll
                        for r in range(1, 24):
                            ws.Rows(page_start_row + r - 1).RowHeight = ws.Rows(r).RowHeight
                        ws.HPageBreaks.Add(ws.Range(f"A{page_start_row}"))
                        
                    # 2단 배치 (0~19는 왼쪽, 20~39는 오른쪽)
                    if pos < 20:
                        row_idx = page_start_row + 3 + pos
                        col_offset = 0
                    else:
                        row_idx = page_start_row + 3 + (pos - 20)
                        col_offset = 6
                        
                    # ORI, RE, TOTAL 계산
                    ori = int(float(item["Film"])) if str(item["Film"]).replace('.','',1).isdigit() else 0
                    re_count = int(float(item["Repair"])) if str(item["Repair"]).replace('.','',1).isdigit() else 0
                    total_count = ori + re_count
                    
                    ws.Cells(row_idx, 1 + col_offset).Value = item["Report No."]
                    ws.Cells(row_idx, 2 + col_offset).Value = item["BOX NO."]
                    ws.Cells(row_idx, 3 + col_offset).Value = item["Joint"]
                    ws.Cells(row_idx, 4 + col_offset).Value = ori if ori > 0 else ""
                    ws.Cells(row_idx, 5 + col_offset).Value = re_count if re_count > 0 else ""
                    ws.Cells(row_idx, 6 + col_offset).Value = total_count if total_count > 0 else ""
                    
                if list_data:
                    max_page = (len(list_data) - 1) // 40
                    last_row = (max_page + 1) * 23
                    ws.PageSetup.PrintArea = f"A1:L{last_row}"
                    
                wb.SaveAs(out_path)
                wb.Close()
                excel.Quit()
                self.add_log(f"📋 외부 템플릿 기반 리스트 엑셀 생성 완료: {os.path.basename(out_path)}")
                return
            except Exception as e:
                self.add_log(f"⚠️ 외부 템플릿 리스트 생성 실패 (기본 양식으로 대체): {e}")
                if 'wb' in locals() and wb: wb.Close(False)
                if 'excel' in locals() and excel: excel.Quit()
                
        # 기본 템플릿 사용
        out_path = os.path.join(save_dir, out_name)
        try:
            # openpyxl을 통해 서식을 예쁘게 적용
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Report List"
            
            headers = ["No.", "BOX NO.", "Report No.", "Joint", "Film", "Repair", "Remark"]
            ws.append(headers)
            
            # 헤더 스타일
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            header_font = Font(bold=True)
            align_center = Alignment(horizontal="center", vertical="center")
            border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for col_idx, col_name in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                cell.border = border_thin
                
            # 데이터 채우기
            for r_idx, row in enumerate(list_data, 2):
                ws.cell(row=r_idx, column=1, value=row["No."])
                ws.cell(row=r_idx, column=2, value=row["BOX NO."])
                ws.cell(row=r_idx, column=3, value=row["Report No."])
                ws.cell(row=r_idx, column=4, value=row["Joint"])
                ws.cell(row=r_idx, column=5, value=float(row["Film"]) if row["Film"].isdigit() else row["Film"])
                ws.cell(row=r_idx, column=6, value=float(row["Repair"]) if row["Repair"].isdigit() else row["Repair"])
                ws.cell(row=r_idx, column=7, value="")
                
                for c_idx in range(1, 8):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.alignment = align_center
                    cell.border = border_thin
                    
            # 열 너비 조정
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 20
            
            wb.save(out_path)
            self.add_log(f"📋 리스트 엑셀 생성 완료: {out_name}")
        except Exception as e:
            # openpyxl 실패 시 pandas 기본 저장으로 대체
            df_list.to_excel(out_path, index=False)
            self.add_log(f"📋 리스트 기본 엑셀 생성 완료: {out_name} (서식 오류: {str(e)})")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMergerApp(root)
    root.mainloop()
