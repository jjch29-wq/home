import openpyxl

wb = openpyxl.load_workbook(r'C:/Users/jjch2/Desktop/템플릿_최종완성본_V74.xlsx', data_only=True)
ws = wb.worksheets[0]

# Scan the whole sheet for NDT result section
print("=== 2. 비파괴검사결과서 섹션 탐색 ===")
for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            v = cell.value.strip()
            if '2.' in v and ('비파괴' in v or 'NDT' in v or '검사결과' in v):
                print(f"Row {cell.row}, Col {cell.column}: {repr(v)}")

print()
print("=== 검사결과 관련 헤더 후보 (업체/라인번호/Joint 있는 행) ===")
for row in ws.iter_rows():
    row_vals = {}
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            cv = cell.value.strip().replace('\n', ' ')
            if cv:
                row_vals[cell.column] = cv

    has_company = any('업체' in v for v in row_vals.values())
    has_joint = any('Joint' in v or 'Joint No' in v for v in row_vals.values())
    has_method = any('검사방법' in v for v in row_vals.values())

    if has_company and has_joint:
        print(f"Row {list(row_vals.keys())[0]//1}: {row_vals}")
