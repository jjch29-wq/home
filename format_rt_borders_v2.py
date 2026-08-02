import openpyxl
from openpyxl.styles import Border, Side

wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')

if 'RT' in wb.sheetnames:
    ws = wb['RT']
    
    dashed_side = Side(style='dashed')
    thin_side = Side(style='thin')
    
    min_r = 3
    max_r = ws.max_row
    min_c = 1
    max_c = 22
    
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            cell = ws.cell(row=r, column=c)
            
            # Default inside borders: vertical dashed, horizontal dashed
            # Let's make internal vertical lines thin (solid) or dashed? 
            # Usually people just use dashed for all internal if they did it quickly, 
            # but let's make horizontal dashed and vertical dashed for inside, 
            # and solid thin for outside.
            
            left = dashed_side
            right = dashed_side
            top = dashed_side
            bottom = dashed_side
            
            if c == min_c: left = thin_side
            if c == max_c: right = thin_side
            if r == min_r: top = thin_side
            if r == max_r: bottom = thin_side
            
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

wb.save(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
print("Outer borders set to solid, inner to dashed.")
