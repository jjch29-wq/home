import openpyxl

template_path = r'C:/Users/jjch2/Desktop/템플릿_최종완성본_V74.xlsx'
wb = openpyxl.load_workbook(template_path)
ws = wb.worksheets[0]

# Write the tags to the template exactly where the user put them in the output file
ws.cell(row=405, column=2).value = '[[NDT_121_PAUT]]'
ws.cell(row=450, column=2).value = '[[NDT_RESULT_PAUT]]'

wb.save(template_path)
print("Tags successfully inserted into the template file!")
