import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\템플릿_최종완성본_V70.xlsx', data_only=True)
ws = wb.active

with open('excel_dump2.txt', 'w', encoding='utf-8') as f:
    r = 382
    row_vals = []
    for c in range(1, 25):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            row_vals.append(str(val).strip().replace('\n', ' '))
    f.write(f"Row {r}: {' | '.join(row_vals)}\n")
    
    r = 403
    row_vals = []
    for c in range(1, 25):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            row_vals.append(str(val).strip().replace('\n', ' '))
    f.write(f"Row {r}: {' | '.join(row_vals)}\n")
