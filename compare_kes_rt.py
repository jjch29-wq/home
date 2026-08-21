import json,re,sys
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

def norm(v):
    if v is None:return ''
    if isinstance(v,float) and v.is_integer():v=int(v)
    return re.sub(r'\s+','',str(v)).upper()
def nj(v):
    s=norm(v);m=re.fullmatch(r'0*(\d+)(.*)',s)
    return str(int(m.group(1)))+m.group(2) if m else s
def nv(v):
    if isinstance(v,(int,float)) and not isinstance(v,bool):return float(v)
    try:return float(str(v).strip())
    except:return 0.0
def extract(p,s):
    ws=load_workbook(p,data_only=True)[s];out=[]
    for r in range(5,ws.max_row+1):
        rep=norm(ws.cell(r,2).value)
        if '-RT-' not in rep:continue
        m=re.search(r'-RT-(\d+)',rep)
        x={'row':r,'report':rep,'report_num':int(m.group(1)) if m else None,'sno':norm(ws.cell(r,3).value),'ident':norm(ws.cell(r,4).value),'joint':nj(ws.cell(r,7).value),'size':norm(ws.cell(r,8).value),'thk':norm(ws.cell(r,9).value),'matl':norm(ws.cell(r,10).value),'welder':norm(ws.cell(r,11).value),'film':nv(ws.cell(r,12).value),'acc':nv(ws.cell(r,13).value),'rej':nv(ws.cell(r,14).value),'rs':nv(ws.cell(r,15).value),'defect':norm(ws.cell(r,16).value),'remark1':norm(ws.cell(r,20).value),'remark2':norm(ws.cell(r,21).value)}
        x['key']=(x['report'],x['ident'],x['joint']);out.append(x)
    return out

a=extract(sys.argv[1],'케이엔솔RT List1');b=extract(sys.argv[2],'RT List')
ma=defaultdict(list);mb=defaultdict(list)
for x in a:ma[x['key']].append(x)
for x in b:mb[x['key']].append(x)
only_a=[];mismatch=[];matched=0
for k,aa in ma.items():
    bb=mb.get(k,[])
    if not bb:only_a+=aa;continue
    matched+=min(len(aa),len(bb));x,y=aa[0],bb[0];d={}
    for f in ['size','thk','matl','welder','film','acc','rej','rs']:
        if x[f]!=y[f]:d[f]={'list1':x[f],'status':y[f]}
    if d:mismatch.append({'key':list(k),'list1_row':x['row'],'status_row':y['row'],'list1_sno':x['sno'],'status_sno':y['sno'],'diffs':d})

nums=[x['report_num'] for x in a if x['report_num'] is not None];lo=min(nums);hi=max(nums)
ka=set(ma);only_b_range=[x for x in b if lo<=x['report_num']<=hi and x['key'] not in ka]

def ag(rows):
 d=defaultdict(lambda:{'rows':0,'film':0,'acc':0,'rej':0,'rs':0})
 for x in rows:
  z=d[x['report']];z['rows']+=1
  for f in ['film','acc','rej','rs']:z[f]+=x[f]
 return d
aa,bb=ag(a),ag(b);reports=[]
for rep in sorted(set(aa)|{x['report'] for x in only_b_range}):
 reports.append({'report':rep,'list1':aa.get(rep,{}),'status':bb.get(rep,{}),'delta_film':aa.get(rep,{}).get('film',0)-bb.get(rep,{}).get('film',0)})

payload={'summary':{'list1_rows':len(a),'list1_film':sum(x['film'] for x in a),'list1_acc':sum(x['acc'] for x in a),'list1_rej':sum(x['rej'] for x in a),'status_all_rows':len(b),'status_all_film':sum(x['film'] for x in b),'matched':matched,'only_list1':len(only_a),'mismatch':len(mismatch),'range_start':lo,'range_end':hi,'only_status_in_range':len(only_b_range),'only_status_in_range_film':sum(x['film'] for x in only_b_range)},'reports':reports,'only_list1':only_a,'mismatch':mismatch,'only_status_in_range':only_b_range}
Path(sys.argv[3]).write_text(json.dumps(payload,ensure_ascii=False,indent=2),encoding='utf-8')
