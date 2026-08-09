import codecs

file_path = r'c:\Users\jjch2\Desktop\PMI\home\src\한국지역난방 중앙지사.py'

with codecs.open(file_path, 'r', 'utf-8') as f:
    lines = f.readlines()

start_idx = -1
end_idx = -1

for i, line in enumerate(lines):
    if 'def write_dynamic_ndt_table(ws, data, title_search):' in line:
        start_idx = i
        break

if start_idx != -1:
    for i in range(start_idx + 1, len(lines)):
        if 'def export_monthly_ndt_report' in line or ('def ' in lines[i] and '    def ' not in lines[i] and '        def ' not in lines[i]):
            # Wait, write_dynamic_ndt_table is INSIDE export_monthly_ndt_report. 
            # It's an inner function.
            pass
        if 'def export_monthly_ndt_report' in lines[i]:
            end_idx = i
            break
        # Let's just find the end by checking indentation. 
        # write_dynamic_ndt_table is indented by 16 spaces.
        if lines[i].strip() != '' and not lines[i].startswith(' ' * 17) and not lines[i].startswith(' ' * 20) and not lines[i].startswith(' ' * 24):
            if lines[i].startswith(' ' * 16) and 'def ' not in lines[i]:
                pass
            if lines[i].startswith(' ' * 16) and 'def ' in lines[i] and 'def write' not in lines[i]:
                end_idx = i
                break
            if len(lines[i]) - len(lines[i].lstrip()) < 16:
                end_idx = i
                break

if start_idx != -1 and end_idx != -1:
    print(f"Found function from {start_idx} to {end_idx}")
    
    new_func = """                def write_dynamic_ndt_table(ws, data, title_search):
                    if not data:
                        return
                        
                    # 1. Find the starting row of the table
                    start_row = None
                    for r in range(300, 10000):
                        val1 = str(ws.cell(row=r, column=2).value or '')
                        val2 = str(ws.cell(row=r, column=3).value or '')
                        if title_search in val1 or title_search in val2:
                            start_row = r + 3
                            break
                            
                    if not start_row:
                        return
                        
                    orig_table_hdr_start = start_row - 3
                    
                    # Sort data by company
                    data = sorted(data, key=lambda x: x.get('company', ''))
                    
                    import openpyxl
                    from copy import copy
                    
                    # Capture formats BEFORE inserting rows
                    # start_row is dummy data row, start_row + 1 is TOTAL row
                    data_styles = {}
                    total_styles = {}
                    for col in range(2, 24):
                        d_cell = ws.cell(row=start_row, column=col)
                        t_cell = ws.cell(row=start_row + 1, column=col)
                        data_styles[col] = {
                            'style': copy(d_cell._style) if d_cell.has_style else None,
                            'font': copy(d_cell.font),
                            'border': copy(d_cell.border),
                            'fill': copy(d_cell.fill),
                            'alignment': copy(d_cell.alignment)
                        }
                        total_styles[col] = {
                            'style': copy(t_cell._style) if t_cell.has_style else None,
                            'font': copy(t_cell.font),
                            'border': copy(t_cell.border),
                            'fill': copy(t_cell.fill),
                            'alignment': copy(t_cell.alignment)
                        }
                    data_height = ws.row_dimensions[start_row].height
                    total_height = ws.row_dimensions[start_row + 1].height
                    
                    def apply_format(r_idx, is_total):
                        styles = total_styles if is_total else data_styles
                        for col, s in styles.items():
                            cell = ws.cell(row=r_idx, column=col)
                            if s['style']: cell._style = copy(s['style'])
                            cell.font = copy(s['font'])
                            cell.border = copy(s['border'])
                            cell.fill = copy(s['fill'])
                            cell.alignment = copy(s['alignment'])
                        ws.row_dimensions[r_idx].height = total_height if is_total else data_height

                    tot_ori, tot_re, tot_total = 0, 0, 0
                    sub_ori, sub_re, sub_total = 0, 0, 0
                    prev_company = None
                    first_row = True
                    
                    for i, row_data in enumerate(data):
                        curr_company = row_data.get('company', '')
                        
                        # Pagination check
                        if pagination_state['available'] <= 0:
                            copy_rows_openpyxl(ws, 374, 379, start_row)
                            copy_rows_openpyxl(ws, orig_table_hdr_start, orig_table_hdr_start + 2, start_row + 6)
                            try: ws.row_breaks.append(openpyxl.worksheet.pagebreak.Break(id=start_row))
                            except Exception: pass
                            start_row += 9
                            pagination_state['available'] = 35
                            
                        # Subtotal Logic
                        if prev_company is not None and curr_company != prev_company:
                            if not first_row: ws.insert_rows(start_row)
                            apply_format(start_row, is_total=True)
                            ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=16)
                            ws.merge_cells(start_row=start_row, start_column=18, end_row=start_row, end_column=19)
                            ws.merge_cells(start_row=start_row, start_column=20, end_row=start_row, end_column=21)
                            ws.merge_cells(start_row=start_row, start_column=22, end_row=start_row, end_column=23)
                            
                            ws.cell(row=start_row, column=2, value=f"{prev_company} 소계")
                            ws.cell(row=start_row, column=17, value=round(sub_ori, 4) if sub_ori else '-')
                            ws.cell(row=start_row, column=18, value=round(sub_re, 4) if sub_re else '-')
                            ws.cell(row=start_row, column=20, value=round(sub_total, 4) if sub_total else '-')
                            
                            start_row += 1
                            pagination_state['available'] -= 1
                            sub_ori, sub_re, sub_total = 0, 0, 0
                            
                        # Data Row Logic
                        if not first_row:
                            ws.insert_rows(start_row)
                        apply_format(start_row, is_total=False)
                        
                        ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row, end_column=5)
                        ws.merge_cells(start_row=start_row, start_column=6, end_row=start_row, end_column=9)
                        ws.merge_cells(start_row=start_row, start_column=10, end_row=start_row, end_column=11)
                        ws.merge_cells(start_row=start_row, start_column=12, end_row=start_row, end_column=13)
                        ws.merge_cells(start_row=start_row, start_column=14, end_row=start_row, end_column=15)
                        ws.merge_cells(start_row=start_row, start_column=18, end_row=start_row, end_column=19)
                        ws.merge_cells(start_row=start_row, start_column=20, end_row=start_row, end_column=21)
                        ws.merge_cells(start_row=start_row, start_column=22, end_row=start_row, end_column=23)
                        
                        ws.cell(row=start_row, column=2, value=curr_company)
                        ws.cell(row=start_row, column=3, value=row_data['seq'])
                        ws.cell(row=start_row, column=4, value=row_data['section'])
                        ws.cell(row=start_row, column=6, value=row_data['line_no'])
                        ws.cell(row=start_row, column=10, value=row_data['inch'])
                        ws.cell(row=start_row, column=12, value=row_data['joints'])
                        ws.cell(row=start_row, column=14, value=row_data['spec'])
                        ws.cell(row=start_row, column=16, value=row_data['unit'])
                        
                        ws.cell(row=start_row, column=17, value=round(row_data['ori'], 4) if row_data['ori'] else '-')
                        ws.cell(row=start_row, column=18, value=round(row_data['re'], 4) if row_data['re'] else '-')
                        ws.cell(row=start_row, column=20, value=round(row_data['total'], 4) if row_data['total'] else '-')
                        ws.cell(row=start_row, column=22, value=row_data['remark'])
                        
                        sub_ori += row_data['ori']
                        sub_re += row_data['re']
                        sub_total += row_data['total']
                        
                        tot_ori += row_data['ori']
                        tot_re += row_data['re']
                        tot_total += row_data['total']
                        
                        prev_company = curr_company
                        start_row += 1
                        if not first_row:
                            pagination_state['available'] -= 1
                        first_row = False
                        
                    # Last Subtotal
                    if prev_company is not None:
                        if pagination_state['available'] <= 0:
                            copy_rows_openpyxl(ws, 374, 379, start_row)
                            copy_rows_openpyxl(ws, orig_table_hdr_start, orig_table_hdr_start + 2, start_row + 6)
                            try: ws.row_breaks.append(openpyxl.worksheet.pagebreak.Break(id=start_row))
                            except Exception: pass
                            start_row += 9
                            pagination_state['available'] = 35
                            
                        ws.insert_rows(start_row)
                        apply_format(start_row, is_total=True)
                        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=16)
                        ws.merge_cells(start_row=start_row, start_column=18, end_row=start_row, end_column=19)
                        ws.merge_cells(start_row=start_row, start_column=20, end_row=start_row, end_column=21)
                        ws.merge_cells(start_row=start_row, start_column=22, end_row=start_row, end_column=23)
                        
                        ws.cell(row=start_row, column=2, value=f"{prev_company} 소계")
                        ws.cell(row=start_row, column=17, value=round(sub_ori, 4) if sub_ori else '-')
                        ws.cell(row=start_row, column=18, value=round(sub_re, 4) if sub_re else '-')
                        ws.cell(row=start_row, column=20, value=round(sub_total, 4) if sub_total else '-')
                        
                        start_row += 1
                        pagination_state['available'] -= 1

                    # Grand Total
                    # original TOTAL row is now exactly at start_row
                    ws.cell(row=start_row, column=2, value="전체 총계(TOTAL)")
                    ws.cell(row=start_row, column=17, value=round(tot_ori, 4) if tot_ori else '-')
                    ws.cell(row=start_row, column=18, value=round(tot_re, 4) if tot_re else '-')
                    ws.cell(row=start_row, column=20, value=round(tot_total, 4) if tot_total else '-')
"""
    
    new_lines = lines[:start_idx] + [new_func + '\n'] + lines[end_idx:]
    with codecs.open(file_path, 'w', 'utf-8') as f:
        f.writelines(new_lines)
    print("Successfully replaced.")
else:
    print("Could not find the bounds.")
