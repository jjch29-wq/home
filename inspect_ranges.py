import json, sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

p=sys.argv[1]
wf=load_workbook(p,data_only=False)
wv=load_workbook(p,data_only=True)
targets={
 '사전원가':(1,104,1,15),
 '사후원가':(1,77,1,13),
 '대비표':(1,20,1,9),
 '인원투입계획':(1,21,1,26),
 '경비예상지출항목':(1,24,1,4),
 '개인일비':(1,24,1,9),
 '장비감각상각':(1,25,1,10),
 '장비감각상각 (2)':(1,23,1,10),
}
out={}
for sn,(r1,r2,c1,c2) in targets.items():
    ws=wf[sn]; vs=wv[sn]; rows=[]
    for r in range(r1,r2+1):
        cells=[]
        for c in range(c1,c2+1):
            f=ws.cell(r,c).value; v=vs.cell(r,c).value
            if f is not None or v is not None:
                cells.append({'cell':f'{get_column_letter(c)}{r}','f':f,'v':v})
        if cells: rows.append({'row':r,'cells':cells})
    out[sn]=rows
open(sys.argv[2],'w',encoding='utf-8').write(json.dumps(out,ensure_ascii=False,indent=2,default=str))
