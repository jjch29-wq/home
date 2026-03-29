import pandas as pd
import os
import glob
import sys
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from tkinter import filedialog, Tk  # 윈도우 창 기능 추가
import re

# 경고 메시지 무시
warnings.simplefilter("ignore")

# ========================================================
# [설정] SCH → 두께(mm) 변환 테이블
# ========================================================
SCH_TO_THK = {
    "1/2": {"5S": 1.65, "10S": 2.11, "40": 2.77, "80": 3.73, "160": 4.75, "XXS": 7.47},
    "3/4": {"5S": 1.65, "10S": 2.11, "40": 2.87, "80": 3.91, "160": 5.56, "XXS": 7.82},
    "1": {"5S": 1.65, "10S": 2.77, "40": 3.38, "80": 4.55, "160": 6.35, "XXS": 9.09},
    "1-1/4": {"5S": 1.65, "10S": 2.77, "40": 3.56, "80": 4.85, "160": 6.35, "XXS": 9.70},
    "1-1/2": {"5S": 1.65, "10S": 2.77, "40": 3.68, "80": 5.08, "160": 7.14, "XXS": 10.16},
    "2": {"5S": 1.65, "10S": 2.77, "40": 3.91, "80": 5.54, "160": 8.74, "XXS": 11.07},
    "2-1/2": {"5S": 2.11, "10S": 3.05, "40": 5.16, "80": 7.01, "160": 9.53, "XXS": 14.02},
    "3": {"5S": 2.11, "10S": 3.05, "40": 5.49, "80": 7.62, "160": 11.13, "XXS": 15.24},
    "4": {"5S": 2.11, "10S": 3.05, "40": 6.02, "80": 8.56, "160": 13.49, "XXS": 17.12},
    "5": {"5S": 2.77, "10S": 3.40, "40": 6.55, "80": 9.53, "120": 12.70, "160": 15.88},
    "6": {"5S": 2.77, "10S": 3.40, "40": 7.11, "80": 10.97, "120": 14.27, "160": 18.26, "XXS": 21.95},
    "8": {"5S": 2.77, "10S": 3.76, "20": 6.35, "30": 7.04, "40": 8.18, "60": 10.31, "80": 12.70, "100": 15.09, "120": 18.26, "140": 20.62, "160": 23.01, "XXS": 22.23},
    "10": {"5S": 3.40, "10S": 4.19, "20": 6.35, "30": 7.80, "40": 9.27, "60": 12.70, "80": 15.09, "100": 18.26, "120": 21.44, "140": 25.40, "160": 28.58},
    "12": {"5S": 3.96, "10S": 4.57, "20": 6.35, "30": 8.38, "40": 10.31, "60": 14.27, "80": 17.48, "100": 21.44, "120": 25.40, "140": 28.58, "160": 33.32},
    "14": {"5S": 3.96, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 9.53, "40": 11.13, "60": 15.09, "80": 19.05, "100": 23.83, "120": 27.79, "140": 31.75, "160": 35.71},
    "16": {"5S": 4.19, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 9.53, "40": 12.70, "60": 16.66, "80": 21.44, "100": 26.19, "120": 30.96, "140": 36.53, "160": 40.49},
    "18": {"5S": 4.19, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 11.13, "40": 14.27, "60": 19.05, "80": 23.83, "100": 29.36, "120": 34.93, "140": 39.67, "160": 45.24},
    "20": {"5S": 4.78, "10S": 5.54, "10": 6.35, "20": 9.53, "30": 12.70, "40": 15.09, "60": 20.62, "80": 26.19, "100": 32.54, "120": 38.10, "140": 44.45, "160": 50.01},
    "24": {"5S": 5.54, "10S": 6.35, "10": 6.35, "20": 9.53, "30": 14.27, "40": 17.48, "60": 24.61, "80": 30.96, "100": 38.89, "120": 46.02, "140": 52.37, "160": 59.54},
}

def convert_sch_to_thk(size_val, thk_val):
    """SCH 값을 두께(mm)로 변환"""
    if pd.isna(thk_val) or str(thk_val).strip() == "":
        return ""
    
    thk_str = str(thk_val).strip().upper()
    
    # 이미 숫자면 그대로 반환
    try:
        val = float(thk_str.replace("MM", "").replace("T", "").strip())
        if 0 < val < 100:
            return f"{val:.2f}"
    except:
        pass
    
    # SCH 추출 (예: SCH40, 40S, S/40S, SCH.40 등)
    sch_match = re.search(r'(?:SCH[.\s]?|S/)?(\d+S?|XXS|XS)', thk_str, re.IGNORECASE)
    if not sch_match:
        return thk_str
    
    sch = sch_match.group(1).upper()
    
    # 40S, 80S 등은 40, 80으로 변환 (5S, 10S는 유지)
    if sch.endswith('S') and sch not in ['5S', '10S', 'XXS', 'XS']:
        sch = sch[:-1]  # S 제거: 40S -> 40, 80S -> 80
    
    # Size 정규화
    if pd.isna(size_val) or str(size_val).strip() == "":
        return thk_str
    
    size_str = str(size_val).strip().replace('"', '').replace("'", "")
    size_str = re.sub(r'\s+', '-', size_str)
    
    # 테이블에서 찾기
    if size_str in SCH_TO_THK and sch in SCH_TO_THK[size_str]:
        return f"{SCH_TO_THK[size_str][sch]:.2f}"
    
    # 정수 사이즈로 시도
    try:
        size_int = str(int(float(size_str)))
        if size_int in SCH_TO_THK and sch in SCH_TO_THK[size_int]:
            return f"{SCH_TO_THK[size_int][sch]:.2f}"
    except:
        pass
    
    return thk_str

# ========================================================
# [1] 작업 파일 선택 (윈도우 창)
# ========================================================
print("📂 데이터를 추출할 엑셀 파일을 선택해주세요...")

# 1. 윈도우 창 숨기기
root = Tk()
root.withdraw()

# 2. 파일 선택 팝업 띄우기
selected_file = filedialog.askopenfilename(
    title="데이터를 추출할 엑셀 파일 선택",
    filetypes=[("Excel Files", "*.xls *.xlsx")]
)

# 3. 취소 버튼 눌렀을 때
if not selected_file:
    print("❌ 파일 선택이 취소되었습니다. 프로그램을 종료합니다.")
    sys.exit()

# 4. 선택한 파일이 들어있는 '폴더 경로'만 추출
folder_path = os.path.dirname(selected_file)

# 5. 작업 위치 이동
os.chdir(folder_path)
print(f"✅ 선택된 작업 경로: {folder_path}")
print("-" * 50)

# ========================================================
# [설정] 파일명은 선택한 파일 기반으로 동적 생성
# ========================================================

# ========================================================
# [0] 공통 함수
# ========================================================
def find_col(df, keywords, exclude=[]):
    """원하는 단어가 포함된 컬럼명을 찾는 함수"""
    for col in df.columns:
        col_str = str(col).upper().replace("_", " ").replace("\n", " ").strip()
        if any(k in col_str for k in keywords):
            if any(ex in col_str for ex in exclude): 
                continue
            return col
    return None

def force_two_digit(val):
    try:
        s = str(val).strip()
        f = float(s)
        if f.is_integer(): return f"{int(f):02d}"
        return s
    except: return str(val).strip()

def fix_material_name(t):
    """재질명을 표준 포맷으로 변환"""
    t_str = str(t).upper()
    if t_str == 'NAN': return ""
    t_str = t_str.replace('A312-TP304', 'S/S').replace('A312-304L', 'S/S').replace('A312-305L', 'S/S').replace('A53-B', 'C/S')
    return t_str.replace('C2','C/S').replace('C4','C/S').replace('CS','C/S').replace('S99','S/S').replace('SS','S/S')

def save_and_format(df, full_path, sort_col_name, joint_col_name):
    if df.empty: return False
    
    # 정렬 및 No 생성
    try:
        df = df.sort_values(by=[sort_col_name, joint_col_name])
    except KeyError:
        print(f"⚠️ 정렬 실패: '{sort_col_name}' 컬럼이 없습니다.")
        pass

    df.insert(0, 'No', range(1, len(df) + 1))

    # 빈칸 삽입 (스타일 맞춤)
    try:
        iso_idx = list(df.columns).index(sort_col_name)
        df.insert(iso_idx + 1, "빈칸1", "")
        df.insert(iso_idx + 2, "빈칸2", "")
    except ValueError: pass

    # 저장 시도 (파일 열려있으면 대기)
    while True:
        try:
            df.to_excel(full_path, index=False)
            break
        except PermissionError:
            print(f"\n🚨 [오류] '{os.path.basename(full_path)}' 파일이 켜져 있습니다.")
            input("   파일을 닫고 엔터 키를 누르세요 (저장 재시도)...")
        except Exception as e:
            print(f"\n🚨 [오류] 저장 중 문제 발생: {e}")
            return False
    
    # 서식 적용
    try:
        wb = load_workbook(full_path)
        ws = wb.active
        
        header_row = [c.value for c in ws[1]]
        try: col_iso_idx_xl = header_row.index(sort_col_name) + 1
        except: col_iso_idx_xl = 3 

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # 전체 테두리 및 가운데 정렬
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                # ISO 컬럼 외에는 가운데 정렬
                if cell.column != col_iso_idx_xl:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # ISO 컬럼 병합 (No, 빈칸1, 빈칸2 포함하여 3칸 병합하는 로직으로 추정)
        # 원본 의도: ISO 컬럼 + 빈칸1 + 빈칸2 -> 3칸 병합
        for row in range(2, ws.max_row + 1): # 헤더 제외
            ws.merge_cells(start_row=row, start_column=col_iso_idx_xl, end_row=row, end_column=col_iso_idx_xl + 2)
            # 병합된 셀 가운데 정렬
            ws.cell(row=row, column=col_iso_idx_xl).alignment = Alignment(horizontal='center', vertical='center')

        wb.save(full_path)
        return True
    except Exception as e:
        print(f"\n🚨 [오류] 서식 적용 실패: {e}")
        return False

# ========================================================
# [2] 선택한 파일에서 데이터 추출
# ========================================================
target_file = selected_file
print(f"📂 분석 대상 파일: {os.path.basename(target_file)}")

data_list_mode1 = []

try:
    xls = pd.ExcelFile(target_file)
    
    # ====================================================
    # [3] 데이터 추출
    # ====================================================
    for sheet_name in xls.sheet_names:
        
        # 헤더 행 찾기 (LINE/ISO 와 JOINT/WELD가 동시에 있는 행)
        header_row_idx = None
        try:
            temp_df = pd.read_excel(target_file, sheet_name=sheet_name, header=None, nrows=30)
        except: continue

        for i, row in temp_df.iterrows():
            row_str = " ".join([str(x).upper() for x in row if pd.notna(x)])
            has_line = any(k in row_str for k in ["LINE", "ISO", "DWG", "DRAWING"])
            has_joint = any(k in row_str for k in ["JOINT", "WELD", "용접"])
            
            if has_line and has_joint:
                header_row_idx = i
                break
        
        if header_row_idx is None: continue

        # 데이터 읽기
        df = pd.read_excel(target_file, sheet_name=sheet_name, header=header_row_idx)

        # [자동보정] 두 줄 헤더 병합 (ORIGIN, FACTOR 등이 바로 아랫줄에 있을 때)
        if not df.empty:
            # 바로 아랫줄(인덱스 0) 확인
            row0_vals = [str(x).upper() for x in df.iloc[0] if pd.notna(x)]
            row0_str = " ".join(row0_vals)
            
            # 2줄 헤더의 특징 단어들
            if any(k in row0_str for k in ["ORIGIN", "FACTOR", "WELDER1", "WELDER2"]):
                print(f"   ℹ️ [자동보정] 2줄 헤더 병합 처리 (시트: {sheet_name})")
                new_columns = []
                for col, val in zip(df.columns, df.iloc[0]):
                    c_txt = str(col) if pd.notna(col) and "Unnamed" not in str(col) else ""
                    v_txt = str(val) if pd.notna(val) else ""
                    # 윗줄과 아랫줄을 합쳐서 새 컬럼명 생성
                    new_columns.append(f"{c_txt} {v_txt}".strip())
                
                df.columns = new_columns
                df = df.iloc[1:] # 아랫줄 데이터 제거
                df.reset_index(drop=True, inplace=True)
        
        # [컬럼 찾기]
        src_iso     = find_col(df, ["ISO", "LINE", "DWG", "DRAWING"], exclude=["JOINT", "WELD"]) 
        src_joint   = find_col(df, ["JOINT NO", "JOINT NUMBER"], exclude=["ISO", "LINE", "ITEM"])
        if not src_joint:
            src_joint = find_col(df, ["JOINT", "WELD"], exclude=["ISO", "LINE", "ITEM", "WPS", "REPORT", "DATE", "TYPE"])
        
        src_mat     = find_col(df, ["MAT", "MATERIAL", "재질", "CLASS", "M'TL"])
        src_size    = find_col(df, ["SIZE", "NPS", "DIA", "INCH", "관경", "ORIGIN", "DI ORIGIN"])
        src_welder  = find_col(df, ["WELDER", "ID", "용접사", "WELDER1"])
        src_result  = find_col(df, ["RESULT", "결과", "판정"])
        src_thk     = find_col(df, ["THK", "THICK", "SCH", "두께"])
        src_weld_type = find_col(df, ["WELD TYPE", "WELDTYPE", "W.TYPE", "TYPE"], exclude=["JOINT"])

        # [진단]
        if not src_iso: src_iso = df.columns[0]
        if not src_joint: 
            # Joint 정보가 없으면 의미 없는 시트
            continue 

        # 데이터 정제
        df = df[df[src_joint].astype(str) != str(src_joint)] # 헤더 반복 제거
        df = df.dropna(subset=[src_joint])
        df = df[df[src_joint].astype(str).str.strip() != ""]
        df[src_iso] = df[src_iso].ffill().bfill() # ISO 번호 채우기

        # 합격 데이터만 필터링 (Result 컬럼이 있을 때만)
        if src_result:
            s = df[src_result].astype(str).str.upper().str.strip()
            cond_pass = s.str.contains('ACC') | s.str.contains('합격') | s.str.contains('OK') | s.str.contains('ACCEPT')
            cond_fail = s.str.contains('UNACC') | s.str.contains('REJ') | s.str.contains('FAIL') | s.str.contains('RW')
            # 합격이고 불합격이 아닌 것
            df = df[cond_pass & ~cond_fail]

        if df.empty: 
            continue
        else:
            print(f"   👉 [시트: {sheet_name}] 합격 데이터 {len(df)}건 추출")

        # 값 추출 및 변환
        col_iso_val    = df[src_iso].astype(str).str.strip()
        col_joint_val  = df[src_joint].apply(force_two_digit)
        col_size_val   = df[src_size] if src_size else ""
        col_mat_val    = df[src_mat].apply(fix_material_name) if src_mat else ""
        col_welder_val = df[src_welder] if src_welder else ""
        col_result_val = df[src_result] if src_result else "Acc"
        
        # [추가] SCH를 두께(mm)로 변환
        if src_thk and src_size:
            col_thk_val = df.apply(lambda row: convert_sch_to_thk(row[src_size], row[src_thk]), axis=1)
        elif src_thk:
            col_thk_val = df[src_thk]
        else:
            col_thk_val = ""
        
        col_weld_type_val = df[src_weld_type] if src_weld_type else ""

        # [모드 1 데이터셋 구성]
        df1 = pd.DataFrame()
        df1['ISO Drawing No.'] = col_iso_val
        df1['Joint'] = col_joint_val
        df1['NPS'] = col_size_val
        df1['Thk.'] = col_thk_val
        df1['Material'] = col_mat_val
        df1['Welder'] = col_welder_val
        df1['Weld Type'] = col_weld_type_val
        df1['결과'] = "Acc"
        data_list_mode1.append(df1)

    # ====================================================
    # [4] 저장 및 마무리
    # ====================================================
    print("\n" + "="*50)
    print("💾 엑셀 리포트 생성 중...")
    
    # 출력 파일명 선택 (사용자에게 물어보기)
    base_name = os.path.splitext(os.path.basename(target_file))[0]
    default_name = f"{base_name}."
    
    output_file = filedialog.asksaveasfilename(
        title="저장할 파일명 입력",
        initialdir=folder_path,
        initialfile=default_name,
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    
    if not output_file:
        print("❌ 저장 취소됨. 프로그램을 종료합니다.")
        input("\n엔터 키를 누르면 종료합니다...")
        sys.exit()
    
    filename_mode1 = os.path.basename(output_file)

    # 저장
    if data_list_mode1:
        final_df1 = pd.concat(data_list_mode1, ignore_index=True)
        # 중복 제거 (ISO와 Joint가 같으면)
        final_df1 = final_df1.drop_duplicates(subset=['ISO Drawing No.', 'Joint'], keep='first')
        
        if save_and_format(final_df1, output_file, 'ISO Drawing No.', 'Joint'):
            print(f" ✅ [완료] {filename_mode1}")
    else:
        print(f" ⚠️ [실패] {filename_mode1} - 추출된 데이터가 없습니다.")

    print("🎉 모든 작업 끝!")
    input("\n엔터 키를 누르면 종료합니다...")

except Exception as e:
    import traceback
    traceback.print_exc()
    print("❌ 오류가 발생했습니다.")
    input("\n엔터 키를 누르면 종료합니다...")