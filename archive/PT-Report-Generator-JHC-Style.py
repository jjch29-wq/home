import pandas as pd
import os
import glob
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import warnings

# 경고 메시지 무시
warnings.simplefilter("ignore")

# ========================================================
# [★설정] 작업 폴더 경로 (자동 인식)
# ========================================================
# 이 파이썬 파일이 있는 위치를 작업 폴더로 설정합니다.
if getattr(sys, 'frozen', False):
    # PyInstaller로 만든 exe 실행 파일인 경우
    folder_path = os.path.dirname(sys.executable)
elif '__file__' in globals():
    # 일반 파이썬 스크립트로 실행하는 경우
    folder_path = os.path.dirname(os.path.abspath(__file__))
else:
    # 주피터 노트북이나 그 외 환경
    folder_path = os.getcwd()

print(f"📂 현재 작업 폴더: {folder_path}")

filename_mode1 = "PT_Report_Style_JHC.xlsx"
filename_mode2 = "PT_Report_Style_JHC1.xlsx"

# ========================================================
# [0] 공통 함수
# ========================================================
def find_col(df, keywords, exclude=[]):
    """원하는 단어가 포함된 컬럼명을 찾는 함수"""
    for col in df.columns:
        col_str = str(col).upper().replace("_", " ").replace("\n", " ").strip()
        # 1. 키워드 포함 확인
        if any(k in col_str for k in keywords):
            # 2. 제외 단어 확인
            if any(ex in col_str for ex in exclude): 
                continue
            return col
    return None

def force_two_digit(val):
    try:
        s = str(val).strip()
        f = float(s)
        if f.is_integer(): return f"{int(f):02d}"
        return s
    except: return str(val).strip()

def fix_material_name(t):
    """재질명을 표준 포맷으로 변환"""
    t_str = str(t).upper()
    if t_str == 'NAN': return ""
    t_str = t_str.replace('A312-304L', 'S/S').replace('A312-305L', 'S/S').replace('A53-B', 'C/S')
    return t_str.replace('C2','C/S').replace('C4','C/S').replace('CS','C/S').replace('S99','S/S').replace('SS','S/S')

def save_and_format(df, full_path, sort_col_name, joint_col_name):
    if df.empty: return False
    
    # 정렬 및 No 생성
    df = df.sort_values(by=[sort_col_name, joint_col_name])
    df.insert(0, 'No', range(1, len(df) + 1))

    # 빈칸 삽입
    try:
        iso_idx = list(df.columns).index(sort_col_name)
        df.insert(iso_idx + 1, "빈칸1", "")
        df.insert(iso_idx + 2, "빈칸2", "")
    except ValueError: pass

    # 저장
    try:
        df.to_excel(full_path, index=False)
    except PermissionError:
        print(f"\n🚨 [오류] '{os.path.basename(full_path)}' 파일이 열려 있습니다! 닫고 다시 실행해주세요.")
        return False
    except Exception as e:
        print(f"\n🚨 [오류] 저장 중 문제 발생: {e}")
        return False
    
    # 서식 적용
    try:
        wb = load_workbook(full_path)
        ws = wb.active
        
        header_row = [c.value for c in ws[1]]
        try: col_iso_idx_xl = header_row.index(sort_col_name) + 1
        except: col_iso_idx_xl = 3 

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                if cell.column != col_iso_idx_xl:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in range(1, ws.max_row + 1):
            ws.merge_cells(start_row=row, start_column=col_iso_idx_xl, end_row=row, end_column=col_iso_idx_xl + 2)
            ws.cell(row=row, column=col_iso_idx_xl).alignment = Alignment(horizontal='center', vertical='center')

        wb.save(full_path)
        return True
    except PermissionError:
        print(f"\n🚨 [오류] 서식 적용 중 '{os.path.basename(full_path)}' 파일이 열려버렸습니다.")
        return False

# ========================================================
# [1] 파일 찾기
# ========================================================
all_files = glob.glob(os.path.join(folder_path, "RFI-K1-JHC-PIP-PT*.xls*")) + \
            glob.glob(os.path.join(folder_path, "rfi-K1-JHC-PIP-PT*.xls*"))
input_files = [f for f in all_files if "PT_Report_" not in f and "~$" not in os.path.basename(f)]

if not input_files:
    print("❌ 현재 폴더에 'RFI-K1-JHC-PIP-PT'로 시작하는 엑셀 파일이 없습니다.")
    print(f"👉 확인된 폴더 위치: {folder_path}")
    input("엔터를 누르면 종료합니다...") # 파일 없을 때 바로 꺼짐 방지
    exit()

target_file = max(input_files, key=os.path.getctime)
print(f"📂 작업 파일: {os.path.basename(target_file)}")

data_list_mode1 = []
data_list_mode2 = []

try:
    xls = pd.ExcelFile(target_file)
    
    # ====================================================
    # [2] 데이터 추출
    # ====================================================
    for sheet_name in xls.sheet_names:
        
        # 헤더 행 찾기
        header_row_idx = None
        try:
            temp_df = pd.read_excel(target_file, sheet_name=sheet_name, header=None, nrows=30)
        except: continue

        for i, row in temp_df.iterrows():
            row_str = " ".join([str(x).upper() for x in row if pd.notna(x)])
            has_line = any(k in row_str for k in ["LINE", "ISO", "DWG", "DRAWING"])
            has_joint = any(k in row_str for k in ["JOINT", "WELD", "용접"])
            
            if has_line and has_joint:
                header_row_idx = i
                break
        
        if header_row_idx is None: continue

        # 데이터 읽기
        df = pd.read_excel(target_file, sheet_name=sheet_name, header=header_row_idx)

        # [자동보정] 두 줄 헤더 병합
        if not df.empty:
            row0_vals = [str(x).upper() for x in df.iloc[0] if pd.notna(x)]
            row0_str = " ".join(row0_vals)
            
            if any(k in row0_str for k in ["ORIGIN", "FACTOR", "WELDER1"]):
                print(f"   ℹ️ [자동보정] 두 줄 헤더가 감지되어 합칩니다. (시트: {sheet_name})")
                new_columns = []
                for col, val in zip(df.columns, df.iloc[0]):
                    c_txt = str(col) if pd.notna(col) and "Unnamed" not in str(col) else ""
                    v_txt = str(val) if pd.notna(val) else ""
                    new_columns.append(f"{c_txt} {v_txt}".strip())
                
                df.columns = new_columns
                df = df.iloc[1:] 
                df.reset_index(drop=True, inplace=True)
        
        # [컬럼 찾기 설정]
        src_iso     = find_col(df, ["ISO", "LINE", "DWG", "DRAWING"], exclude=["JOINT", "WELD"]) 
        src_joint   = find_col(df, ["JOINT NO", "JOINT NUMBER"], exclude=["ISO", "LINE", "ITEM"])
        if not src_joint:
            src_joint = find_col(df, ["JOINT", "WELD"], exclude=["ISO", "LINE", "ITEM", "WPS", "REPORT", "DATE", "TYPE"])
        
        src_mat     = find_col(df, ["MAT", "MATERIAL", "재질", "CLASS"])
        src_size    = find_col(df, ["SIZE", "NPS", "DIA", "INCH", "관경", "ORIGIN", "DI ORIGIN"])
        src_welder  = find_col(df, ["WELDER", "ID", "용접사", "WELDER1"])
        src_result  = find_col(df, ["RESULT", "결과", "판정"])
        src_thk     = find_col(df, ["THK", "THICK", "SCH", "두께"])

        # [진단용 출력]
        print(f"\n🔍 [시트: {sheet_name}] 컬럼 인식 결과:")
        
        if not src_iso: src_iso = df.columns[0]
        if not src_joint: 
            print("   ⚠️ Joint 컬럼을 못 찾아서 이 시트는 건너뜁니다.")
            continue 

        # 정제
        df = df[df[src_joint].astype(str) != str(src_joint)] 
        df = df.dropna(subset=[src_joint])
        df = df[df[src_joint].astype(str).str.strip() != ""]
        df[src_iso] = df[src_iso].ffill().bfill()

        # 필터링
        if src_result:
            s = df[src_result].astype(str).str.upper().str.strip()
            cond_pass = s.str.contains('ACC') | s.str.contains('합격') | s.str.contains('OK') | s.str.contains('ACCEPT')
            cond_fail = s.str.contains('UNACC') | s.str.contains('REJ') | s.str.contains('FAIL') | s.str.contains('RW')
            df = df[cond_pass & ~cond_fail]

        if df.empty: 
            print("   ⚠️ 필터링 후 남은 데이터가 0개입니다.")
            continue
        else:
            print(f"   👉 {len(df)}개의 합격 데이터를 찾았습니다!")

        # 값 추출
        col_iso_val    = df[src_iso].astype(str).str.strip()
        col_joint_val  = df[src_joint].apply(force_two_digit)
        col_size_val   = df[src_size] if src_size else ""
        col_mat_val    = df[src_mat].apply(fix_material_name) if src_mat else ""
        col_welder_val = df[src_welder] if src_welder else ""
        col_result_val = df[src_result] if src_result else "Acc"
        col_thk_val    = df[src_thk] if src_thk else ""

        # [모드 1]
        df1 = pd.DataFrame()
        df1['ISO Drawing No.'] = col_iso_val
        df1['Joint'] = col_joint_val
        df1['NPS'] = col_size_val
        df1['Thk.'] = col_thk_val
        df1['Material'] = col_mat_val
        df1['Welder'] = col_welder_val
        df1['결과'] = "Acc"
        data_list_mode1.append(df1)

        # [모드 2]
        df2 = pd.DataFrame()
        df2['Iso No.'] = col_iso_val
        df2['Joint No.'] = col_joint_val
        df2['Size (inch)'] = col_size_val
        df2['Thickness'] = col_thk_val
        df2['Material'] = col_mat_val
        df2['용접사 ID'] = col_welder_val
        df2['Result'] = col_result_val
        data_list_mode2.append(df2)

    # ====================================================
    # [3] 저장
    # ====================================================
    print("\n" + "="*50)
    print("💾 엑셀 파일 생성 중...")

    if data_list_mode1:
        final_df1 = pd.concat(data_list_mode1, ignore_index=True)
        final_df1 = final_df1.drop_duplicates(subset=['ISO Drawing No.', 'Joint'], keep='first')
        
        path1 = os.path.join(folder_path, filename_mode1)
        if save_and_format(final_df1, path1, 'ISO Drawing No.', 'Joint'):
            print(f"  ✅ [완료] {filename_mode1}")
    else:
        print("  ⚠️ 데이터가 추출되지 않아 파일을 만들지 못했습니다.")

    if data_list_mode2:
        final_df2 = pd.concat(data_list_mode2, ignore_index=True)
        final_df2 = final_df2.drop_duplicates(subset=['Iso No.', 'Joint No.'], keep='first')
        
        path2 = os.path.join(folder_path, filename_mode2)
        if save_and_format(final_df2, path2, 'Iso No.', 'Joint No.'):
            print(f"  ✅ [완료] {filename_mode2}")

    print("🎉 모든 작업 끝!")
    # 결과 확인을 위해 잠시 멈춤 (엔터 누르면 꺼짐)
    input("\n엔터 키를 누르면 종료합니다...")

except Exception as e:
    import traceback
    traceback.print_exc()
    print("❌ 오류가 발생했습니다.")
    input("\n엔터 키를 누르면 종료합니다...")