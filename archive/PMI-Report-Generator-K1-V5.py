import sys
import os
import glob
import math
import traceback
import re
import warnings
import json
import tempfile
import datetime
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, Border, Side
from PIL import Image as PILImage, ImageChops
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# Ignore library warnings
warnings.simplefilter("ignore")

# --- Constants & Paths ---
NAN_PATTERN = re.compile(r'^nan(\.0+)?$|^none$|^null$|^0\.0+|-0\.0+$', re.IGNORECASE)
DOT_ZERO_PATTERN = re.compile(r'\.0$')

# [NEW] 공통 스타일 (테두리 등)
thin_side = Side(style='thin')
medium_side = Side(style='medium')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

if getattr(sys, 'frozen', False):
    SCRIPT_HOME = os.path.dirname(sys.executable)
else:
    SCRIPT_HOME = os.path.dirname(os.path.abspath(__file__))

SETTINGS_FILE = os.path.join(SCRIPT_HOME, "logo_settings.json")

class PMIReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PMI 성적서 자동 생성기 (Professional v2.0)")
        self.root.geometry("800x850")
        self.root.configure(background="#f9fafb")
        
        # --- State Variables ---
        self.logo_folder_path = tk.StringVar(value=SCRIPT_HOME)
        self.target_file_path = tk.StringVar()
        self.template_file_path = tk.StringVar()
        self.sequence_filter = tk.StringVar()
        
        # Default Settings (will be overwritten by load_settings)
        self.config = {
            # 갑지 (Cover)
            'SEOUL_COVER_ANCHOR': "E5", 'SEOUL_COVER_W': 200.0, 'SEOUL_COVER_H': 18.0, 'SEOUL_COVER_X': 30.0, 'SEOUL_COVER_Y': 15.0,
            'SITCO_COVER_ANCHOR': "A6", 'SITCO_COVER_W': 80.0, 'SITCO_COVER_H': 40.0, 'SITCO_COVER_X': 15.0, 'SITCO_COVER_Y': 10.0,
            'FOOTER_COVER_ANCHOR': "Q51", 'FOOTER_COVER_W': 80.0, 'FOOTER_COVER_H': 20.0, 'FOOTER_COVER_X': -10.0, 'FOOTER_COVER_Y': 5.0,
            'FOOTER_PT_COVER_ANCHOR': "A51", 'FOOTER_PT_COVER_W': 100.0, 'FOOTER_PT_COVER_H': 25.0, 'FOOTER_PT_COVER_X': 3.0, 'FOOTER_PT_COVER_Y': 5.0,
            # 을지 (Data)
            'SEOUL_DATA_ANCHOR': "F5", 'SEOUL_DATA_W': 200.0, 'SEOUL_DATA_H': 18.0, 'SEOUL_DATA_X': 35.0, 'SEOUL_DATA_Y': 15.0,
            'SITCO_DATA_ANCHOR': "A6", 'SITCO_DATA_W': 80.0, 'SITCO_DATA_H': 40.0, 'SITCO_DATA_X': 5.0, 'SITCO_DATA_Y': 10.0,
            'FOOTER_DATA_ANCHOR': "Q37", 'FOOTER_DATA_W': 100.0, 'FOOTER_DATA_H': 15.0, 'FOOTER_DATA_X': 5.0, 'FOOTER_DATA_Y': 3.0,
            'FOOTER_PT_DATA_ANCHOR': "A37", 'FOOTER_PT_DATA_W': 100.0, 'FOOTER_PT_DATA_H': 30.0, 'FOOTER_PT_DATA_X': 3.0, 'FOOTER_PT_DATA_Y': -10.0,
            # 행 설정
            'START_ROW': 19, 'DATA_END_ROW': 45, 'PRINT_END_ROW': 47
        }
        
        self.load_settings()
        
        # [FIX] Migration: Force row settings to 19-46 range
        migrated = False
        if self.config.get('START_ROW') == 20:
            self.config['START_ROW'] = 19; migrated = True
        if self.config.get('DATA_END_ROW') in [46, 48, 51]:
            self.config['DATA_END_ROW'] = 45; migrated = True
        if self.config.get('PRINT_END_ROW') in [48, 50, 53]:
            self.config['PRINT_END_ROW'] = 47; migrated = True
            
        if migrated: self.save_settings()

        self.create_widgets()
        self.log("[INFO] 프로그램을 시작했습니다.")
        
    def log(self, message):
        """Append message to status log"""
        if hasattr(self, 'status_log'):
            self.status_log.config(state='normal')
            self.status_log.insert(tk.END, f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}\n")
            self.status_log.see(tk.END)
            self.status_log.config(state='disabled')
            self.root.update_idletasks()
        else:
            print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}")

    def load_settings(self):
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    saved_data = json.load(f)
                    self.config.update(saved_data)
                print("SUCCESS: 저장된 설정을 불러왔습니다.")
            except Exception as e:
                print(f"WARNING: 설정 불러오기 실패 (기본값 사용): {e}")

    def save_settings(self):
        try:
            # Sync settings from UI entries if they exist
            if hasattr(self, 'setting_vars'):
                for key, var in self.setting_vars.items():
                    val = var.get()
                    try:
                        if key.endswith(('_X', '_Y', '_W', '_H')): self.config[key] = float(val)
                        elif 'ROW' in key: self.config[key] = int(val)
                        else: self.config[key] = str(val)
                    except: pass

            with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=4)
            self.log("[SUCCESS] 설정이 파일에 저장되었습니다.")
        except Exception as e:
            self.log(f"[WARNING] 설정 저장 실패: {e}")

    def create_widgets(self):
        # --- Main Style ---
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TFrame", background="#f9fafb")
        style.configure("TLabel", background="#f9fafb", font=("Malgun Gothic", 10))
        style.configure("TLabelframe", background="#f9fafb", font=("Malgun Gothic", 10, "bold"))
        style.configure("TLabelframe.Label", background="#f9fafb", font=("Malgun Gothic", 10, "bold"))
        style.configure("Action.TButton", font=("Malgun Gothic", 11, "bold"), padding=10)

        main_container = tk.Frame(self.root, background="#f9fafb", padx=20, pady=20)
        main_container.pack(fill='both', expand=True)

        # Header Title
        tk.Label(main_container, text="PMI 성적서 생성 대시보드", font=("Malgun Gothic", 16, "bold"), background="#f9fafb", foreground="#111827").pack(pady=(0, 20), anchor='w')

        # 1. File Selection Section
        file_frame = ttk.LabelFrame(main_container, text=" 파일 및 폴더 선택 ", padding=15)
        file_frame.pack(fill='x', pady=(0, 20))

        def _add_file_row(parent, label, var, row, is_dir=False, types=None):
            ttk.Label(parent, text=label).grid(row=row, column=0, sticky='e', padx=5, pady=5)
            ttk.Entry(parent, textvariable=var, width=50).grid(row=row, column=1, padx=5, pady=5)
            cmd = (lambda: self._browse_dir(var)) if is_dir else (lambda: self._browse_file(var, types))
            ttk.Button(parent, text="찾기", command=cmd).grid(row=row, column=2, padx=5, pady=5)

        _add_file_row(file_frame, "로고 폴더:", self.logo_folder_path, 0, is_dir=True)
        _add_file_row(file_frame, "RFI 데이터:", self.target_file_path, 1, types=[("Excel Source", "*.xls;*.xlsx;*.xlsm")])
        _add_file_row(file_frame, "성적서 양식:", self.template_file_path, 2, types=[("Excel Template", "*.xlsx;*.xlsm")])

        # 2. Configuration Tabs
        config_frame = ttk.LabelFrame(main_container, text=" 리포트 세부 설정 ", padding=5)
        config_frame.pack(fill='both', expand=False, pady=(0, 20))

        self.tab_notebook = ttk.Notebook(config_frame)
        self.tab_notebook.pack(fill='both', expand=True)

        self.setting_vars = {}
        tab_cover = ttk.Frame(self.tab_notebook, padding=10)
        tab_data = ttk.Frame(self.tab_notebook, padding=10)
        tab_rows = ttk.Frame(self.tab_notebook, padding=10)

        self.tab_notebook.add(tab_cover, text="갑지 (Cover)")
        self.tab_notebook.add(tab_data, text="을지 (Data)")
        self.tab_notebook.add(tab_rows, text="행 설정 (Rows)")

        self._create_setting_grid(tab_cover, "COVER")
        self._create_setting_grid(tab_data, "DATA")
        self._create_row_settings(tab_rows)

        # 3. Action Section
        action_outer = tk.Frame(main_container, background="#f9fafb")
        action_outer.pack(fill='x', pady=(0, 10))

        action_frame = tk.Frame(action_outer, background="#ffffff", highlightthickness=1, highlightbackground="#e5e7eb", padx=15, pady=15)
        action_frame.pack(fill='x')

        ttk.Label(action_frame, text="특정 순번(NO)만 추출 (예: 1, 3, 5-10):", background="#ffffff").pack(side='left', padx=(0, 10))
        ttk.Entry(action_frame, textvariable=self.sequence_filter, width=30).pack(side='left')
        
        btn_start = ttk.Button(action_frame, text="성적서 생성 시작 (Generate)", style="Action.TButton", command=self.run_process)
        btn_start.pack(side='right')

        # 4. Progress bar
        self.progress = ttk.Progressbar(main_container, orient="horizontal", mode="determinate")
        self.progress.pack(fill='x', pady=(5, 15))

        # 5. Status Log Area
        log_frame = ttk.LabelFrame(main_container, text=" 작업 로그 (Status Log) ", padding=10)
        log_frame.pack(fill='both', expand=True)

        self.status_log = tk.Text(log_frame, height=8, font=("Consolas", 9), state='disabled', background="#000000", foreground="#10b981", padx=5, pady=5)
        vsb = ttk.Scrollbar(log_frame, orient="vertical", command=self.status_log.yview)
        self.status_log.configure(yscrollcommand=vsb.set)
        
        self.status_log.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

    def _create_setting_grid(self, parent, context):
        items = [
            ("SITCO 로고", f"SITCO_{context}"),
            ("서울검사 로고", f"SEOUL_{context}"),
            ("바닥글 우측", f"FOOTER_{context}"),
            ("바닥글 좌측", f"FOOTER_PT_{context}")
        ]
        
        for i, (label, key_prefix) in enumerate(items):
            ttk.Label(parent, text=label, font=("Malgun Gothic", 9, "bold")).grid(row=i, column=0, sticky='w', pady=(10, 2))
            
            ttk.Label(parent, text="셀:").grid(row=i, column=1, sticky='e')
            v_a = tk.StringVar(value=self.config[f"{key_prefix}_ANCHOR"])
            ttk.Entry(parent, textvariable=v_a, width=6).grid(row=i, column=2, padx=2)
            self.setting_vars[f"{key_prefix}_ANCHOR"] = v_a
            
            for j, (coord, key_suffix) in enumerate([("X", "X"), ("Y", "Y"), ("W", "W"), ("H", "H")]):
                ttk.Label(parent, text=f"{coord}:").grid(row=i, column=3 + j*2, sticky='e')
                v = tk.StringVar(value=str(self.config[f"{key_prefix}_{key_suffix}"]))
                ttk.Entry(parent, textvariable=v, width=6).grid(row=i, column=4 + j*2, padx=2)
                self.setting_vars[f"{key_prefix}_{key_suffix}"] = v

    def _create_row_settings(self, parent):
        rows = [
            ("데이터 시작 행", "START_ROW", "실제 데이터가 입력되는 첫 행"),
            ("데이터 종료 행", "DATA_END_ROW", "한 페이지에서 데이터 입력이 끝나는 행"),
            ("인쇄 영역 종료 행", "PRINT_END_ROW", "페이지 구분선이 위치할 마지막 행")
        ]
        for i, (label, key, tip) in enumerate(rows):
            ttk.Label(parent, text=label + ":").grid(row=i, column=0, sticky='e', padx=5, pady=8)
            v = tk.StringVar(value=str(self.config[key]))
            ttk.Entry(parent, textvariable=v, width=10).grid(row=i, column=1, sticky='w', padx=5)
            self.setting_vars[key] = v
            ttk.Label(parent, text=f"({tip})", foreground="gray", font=("Malgun Gothic", 9)).grid(row=i, column=2, sticky='w')

    def _browse_dir(self, var):
        path = filedialog.askdirectory(initialdir=var.get() or SCRIPT_HOME)
        if path: var.set(path)

    def _browse_file(self, var, types):
        path = filedialog.askopenfilename(initialdir=os.path.dirname(var.get() or SCRIPT_HOME), filetypes=types)
        if path: var.set(path)

    # --- Excel Helper Logic ---

    def find_image_smart(self, keyword, exclude_keyword=None):
        folder = self.logo_folder_path.get()
        if not folder: return None
        candidates = glob.glob(os.path.join(folder, "*.*"))
        valid_extensions = ['.PNG', '.JPG', '.JPEG', '.BMP', '.GIF']
        for path in candidates:
            fname = os.path.basename(path).upper(); ext = os.path.splitext(path)[1].upper()
            if ext not in valid_extensions: continue 
            if keyword.upper() in fname:
                if exclude_keyword and exclude_keyword.upper() in fname: continue
                return path 
        return None

    def place_image_freely(self, ws, img_path, anchor_cell_str, w, h, x_offset, y_offset):
        try:
            if not img_path or not os.path.exists(img_path): return
            original = PILImage.open(img_path).convert("RGBA")
            resized = original.resize((int(w), int(h)), PILImage.Resampling.LANCZOS)
            temp_name = f"temp_{os.path.basename(img_path)}"
            temp_full_path = os.path.join(tempfile.gettempdir(), temp_name)
            resized.save(temp_full_path)
            
            img = XLImage(temp_full_path); img.width = w; img.height = h
            col_str, row_num = coordinate_from_string(anchor_cell_str)
            col_idx = column_index_from_string(col_str) - 1; row_idx = row_num - 1 
            
            emu_x = int(x_offset * 9525); emu_y = int(y_offset * 9525)
            emu_w = int(w * 9525); emu_h = int(h * 9525)
            
            marker = AnchorMarker(col=col_idx, colOff=emu_x, row=row_idx, rowOff=emu_y)
            size = XDRPositiveSize2D(cx=emu_w, cy=emu_h)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(img)
        except Exception as e: self.log(f"[WARNING] {os.path.basename(img_path)} 배치 실패: {e}")

    def add_logos_to_sheet(self, ws, is_cover=False, clear_existing=True):
        if clear_existing:
            try: ws._images = [] 
            except: pass
        ctx = "COVER" if is_cover else "DATA"
        
        def _get_params(key_prefix):
            return (
                self.config[f"{key_prefix}_{ctx}_ANCHOR"],
                self.config[f"{key_prefix}_{ctx}_W"],
                self.config[f"{key_prefix}_{ctx}_H"],
                self.config[f"{key_prefix}_{ctx}_X"],
                self.config[f"{key_prefix}_{ctx}_Y"]
            )

        path = self.find_image_smart("SITCO"); self.place_image_freely(ws, path, *_get_params("SITCO"))
        path = self.find_image_smart("서울검사"); self.place_image_freely(ws, path, *_get_params("SEOUL"))
        
        path = self.find_image_smart("바닥글", exclude_keyword="PMI-1")
        if not path: path = self.find_image_smart("PMI", exclude_keyword="PMI-1")
        self.place_image_freely(ws, path, *_get_params("FOOTER"))
        
        path_left = self.find_image_smart("PMI갑") if is_cover else None
        if not path_left: path_left = self.find_image_smart("PMI-1")
        if not path_left: path_left = self.find_image_smart("PT")
        self.place_image_freely(ws, path_left, *_get_params("FOOTER_PT"))

    def force_print_settings(self, ws):
        try:
            ws.print_area = f'A1:L{self.config["PRINT_END_ROW"]}'
            ws.row_breaks.brk = []; ws.col_breaks.brk = []
            ws.page_setup.paperSize = 9; ws.page_setup.orientation = 'portrait'
            
            # [ADJUST] Use fixed 95% scale as requested
            ws.page_setup.scale = 95
            ws.page_setup.fitToHeight = None; ws.page_setup.fitToWidth = None
            ws.sheet_properties.pageSetUpPr.fitToPage = False
            
            ws.print_options.horizontalCentered = True; ws.print_options.verticalCentered = True
            
            # [ADJUST] Increase left margin to balance (Left 0.5, Right 0.3)
            ws.page_margins.left = 0.5; ws.page_margins.right = 0.3; ws.page_margins.top = 0.2; ws.page_margins.bottom = 0.2
            ws.page_margins.header = 0.2; ws.page_margins.footer = 0.2
        except: pass

    def set_eulji_headers(self, ws):
        headers = ["NI", "CR", "MO"] # H, I, J열에 표시
        data_font = Font(size=9)
        header_row = self.config['START_ROW']
        
        # 이전 헤더 영역 초기화 (G~L)
        for c in range(7, 13):
            ws.cell(row=header_row, column=c).value = None
            
        for i, val in enumerate(headers):
            col = 8 + i # H=8, I=9, J=10
            cell = ws.cell(row=header_row, column=col)
            cell.value = val; cell.alignment = Alignment(horizontal='center', vertical='center'); cell.font = data_font
        
        # [NEW] Data Validation for Column G (Component Type)
        comp_list = "Tee,Weld,Pipe,Elbow,Cap,Flang"
        dv_g = DataValidation(type="list", formula1=f'"{comp_list}"', allow_blank=True)
        ws.add_data_validation(dv_g)
        
        # [NEW] Data Validation for Column Q (Material Grade)
        materials = "SS304,SS304L,SS316,SS316L,SS321,SS347,SS410,SS430,DUPLEX,MONEL,INCONEL,ER308,ER308L,ER309,ER309L,ER316,ER316L,ER347,ER2209,WP316,WP316L,TP316,TP316L,F316L,A182-F316L,A312-TP316L"
        dv_q = DataValidation(type="list", formula1=f'"{materials}"', allow_blank=True)
        ws.add_data_validation(dv_q)
        
        for r in range(self.config['START_ROW'], self.config['DATA_END_ROW'] + 1):
            # L열 (Grade) - [MOVED from M]
            target_l = ws.cell(row=r, column=12)
            dv_q.add(target_l)
            target_l.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            target_l.font = Font(size=8.5)
            
            # G열 (No / Type)
            target_g = ws.cell(row=r, column=7)
            dv_g.add(target_g)
            target_g.alignment = Alignment(horizontal='center', vertical='center')

    def prepare_next_sheet(self, wb, source_sheet_idx, page_num):
        source_sheet = wb.worksheets[source_sheet_idx]; new_sheet = wb.copy_worksheet(source_sheet) 
        base_title = source_sheet.title.split('_')[0]; new_sheet.title = f"{base_title[:20]}_{page_num:03d}"
        new_sheet.sheet_view.tabSelected = False
        try: new_sheet["K3"] = f"{page_num}"
        except: pass
        self.force_print_settings(new_sheet)
        self.add_logos_to_sheet(new_sheet, is_cover=False)
        
        # Explicitly copy column widths from source to new sheet
        for col_letter, col_dim in source_sheet.column_dimensions.items():
            new_sheet.column_dimensions[col_letter].width = col_dim.width

        data_font = Font(size=9); grade_font = Font(size=8.5)
        for r in range(self.config['START_ROW'], self.config['DATA_END_ROW'] + 1):
            rd = new_sheet.row_dimensions[r]
            rd.height = 20.55 # 27 rows * 1.05pt increase = ~28.35pt (10mm) total increase
            try: rd.custom_height = True
            except: pass
            for c in range(1, 13): # A~L열 초기화
                cell = new_sheet.cell(row=r, column=c)
                cell.font = grade_font if c == 12 else data_font
                if not isinstance(cell, MergedCell): cell.value = None

        # [FIX] 을지 데이터 영역의 기존 병합 해제 (필수)
        merged_to_clear = []
        for rng in new_sheet.merged_cells.ranges:
            if rng.min_row >= self.config['START_ROW'] and rng.max_row <= self.config['DATA_END_ROW']:
                merged_to_clear.append(rng)
        for rng in merged_to_clear:
            new_sheet.unmerge_cells(str(rng))

        return new_sheet

    def run_process(self):
        target_file = self.target_file_path.get()
        template_path = self.template_file_path.get()
        
        if not target_file or not template_path:
            messagebox.showwarning("파일 미선택", "데이터 파일과 양식 파일을 모두 선택해주세요.")
            return

        self.save_settings()
        self.log("🚀 프로세스 시작...")
        self.progress['value'] = 0
        
        try:
            target_input = self.sequence_filter.get().strip()
            target_no_list = [x.strip() for x in target_input.replace(',', ' ').split() if x.strip()] if target_input else []
            
            all_extracted_data = []
            xls = pd.ExcelFile(target_file)
            
            total_sheets = len(xls.sheet_names)
            for s_idx, sheet_name in enumerate(xls.sheet_names):
                self.log(f"📄 시트 처리 중: {sheet_name}")
                try: temp_df = pd.read_excel(target_file, sheet_name=sheet_name, header=None, nrows=50)
                except Exception as e: 
                    self.log(f"⚠️ {sheet_name} 읽기 실패: {e}")
                    continue
                
                header_idx = None
                for i, row in temp_df.iterrows():
                    row_str = str(row.values).upper()
                    if ("CR" in row_str and "NI" in row_str) or ("CHROMIUM" in row_str): header_idx = i; break
                
                if header_idx is None:
                    header_idx = simpledialog.askinteger("헤더 검색 실패", f"'{sheet_name}' 시트의 제목 줄 번호(0부터 시작)를 입력하세요:", parent=self.root)
                    if header_idx is None: continue
                
                df = pd.read_excel(target_file, sheet_name=sheet_name, header=header_idx)
                
                def _find_col(df, keywords):
                    for col in df.columns:
                        if any(k in str(col).upper() for k in keywords): return col
                    return None

                col_cr = _find_col(df, ["CR", "CHROMIUM"]); col_ni = _find_col(df, ["NI", "NICKEL"])
                if not (col_cr and col_ni):
                    self.log(f"[WARNING] {sheet_name}: 필수 컬럼(Cr/Ni) 미발견")
                    continue
                
                col_mo = _find_col(df, ["MO", "MOLYBDENUM"]); col_mn = _find_col(df, ["MN", "MANGANESE"])
                col_no = _find_col(df, ["NO.", "NO", "SEQ", "NUM", "POS", "ITEM"])
                col_dwg = _find_col(df, ["ISO", "DWG", "DRAWING", "LINE"])
                col_grade_orig = _find_col(df, ["GRADE", "MATERIAL", "SPEC", "TYPE"])

                def _to_f(val):
                    if pd.isna(val): return 0.0
                    s = str(val).upper().replace("%", "").strip()
                    if "<" in s or "ND" in s or s == "": return 0.0
                    try: return float(s)
                    except: return 0.0

                count = 0
                for _, row in df.iterrows():
                    v_raw_no = str(row[col_no]).strip() if col_no is not None else str(_+1)
                    if target_no_list and v_raw_no not in target_no_list: continue
                    v_cr = _to_f(row[col_cr])
                    if v_cr > 0 or (v_raw_no != "" and v_raw_no != "nan"):
                        v_mo = _to_f(row[col_mo]) if col_mo is not None else 0.0
                        orig_grade = str(row[col_grade_orig]).strip() if col_grade_orig is not None else ""
                        v_dwg = str(row[col_dwg]).strip() if col_dwg is not None else ""
                        if v_dwg == "nan": v_dwg = ""
                        
                        all_extracted_data.append({
                            'No': v_raw_no, 'Cr': v_cr, 'Ni': _to_f(row[col_ni]),
                            'Mo': v_mo, 'Mn': _to_f(row[col_mn]) if col_mn is not None else 0.0,
                            'Grade': orig_grade if (orig_grade and orig_grade != "nan") else ("SS316" if v_mo >= 1.5 else "SS304"),
                            'Dwg': v_dwg
                        })
                        count += 1
                self.log(f"   -> {count}개 아이템 추출 완료")
                self.progress['value'] = ((s_idx + 1) / total_sheets) * 30

            if not all_extracted_data:
                messagebox.showerror("오류", "추출된 데이터가 없습니다.")
                return

            if target_no_list:
                all_extracted_data.sort(key=lambda x: target_no_list.index(str(x['No'])) if str(x['No']) in target_no_list else 999999)

            self.log("📝 엑셀 성적서 작성 중...")
            wb = openpyxl.load_workbook(template_path, keep_vba=True)
            if hasattr(wb, 'external_links'): wb.external_links.clear()

            # Process 1st sheet (Cover)
            if len(wb.worksheets) >= 1:
                ws0 = wb.worksheets[0]
                self.add_logos_to_sheet(ws0, is_cover=True, clear_existing=False)
                
                # [FIX] Enforce consistent left border for Gapji specifically for rows 23-38
                for r in range(23, 39):
                    # Left border (Column A)
                    cell_a = ws0.cell(row=r, column=1)
                    existing_b = cell_a.border
                    cell_a.border = Border(left=medium_side, right=existing_b.right, top=existing_b.top, bottom=existing_b.bottom)
                    
                    # [NEW] Increase row height by 2 pixels (+1.5pt)
                    rd = ws0.row_dimensions[r]
                    rd.height = 16.5
                    try: rd.custom_height = True
                    except: pass
                
                # [FIX] Remove unwanted border from cell I35
                ws0['I35'].border = Border()
            
            # Process Data Sheets
            data_sheet_id = 1 if len(wb.worksheets) >= 2 else 0
            ws = wb.worksheets[data_sheet_id]
            # First Data Sheet Init
            ws.title = f"{ws.title[:20]}_001"
            
            # Ensure template column widths are respected (some might be lost during open)
            # No changes needed here usually, but good to be aware.

            self.add_logos_to_sheet(ws, is_cover=False)
            self.force_print_settings(ws)
            self.set_eulji_headers(ws)
            for r in range(self.config['START_ROW'], self.config['DATA_END_ROW'] + 1):
                rd = ws.row_dimensions[r]
                rd.height = 20.55
                try: rd.custom_height = True
                except: pass
            # [REMOVED] delete_rows(38, 4) - No longer needed as data goes up to 48

            # [FIX] Define DataValidation objects before the loop
            comp_list = "Tee,Weld,Pipe,Elbow,Cap,Flang"
            dv_g = DataValidation(type="list", formula1=f'"{comp_list}"', allow_blank=True)
            materials = "SS304,SS304L,SS316,SS316L,SS321,SS347,SS410,SS430,DUPLEX,MONEL,INCONEL,ER308,ER308L,ER309,ER309L,ER316,ER316L,ER347,ER2209,WP316,WP316L,TP316,TP316L,F316L,A182-F316L,A312-TP316L"
            dv_q = DataValidation(type="list", formula1=f'"{materials}"', allow_blank=True)
            
            # Add validations to the current (first) sheet
            ws.add_data_validation(dv_g)
            ws.add_data_validation(dv_q)

            # [FIX] 첫 을지 시트의 데이터 영역 병합 해제
            merged_to_clear = []
            for rng in ws.merged_cells.ranges:
                if rng.min_row >= self.config['START_ROW'] and rng.max_row <= self.config['DATA_END_ROW']:
                    merged_to_clear.append(rng)
            for rng in merged_to_clear:
                ws.unmerge_cells(str(rng))

            current_row = self.config['START_ROW']
            current_page = 1
            data_ptr = 0
            
            while data_ptr < len(all_extracted_data):
                if current_row > self.config['DATA_END_ROW']:
                    current_page += 1
                    ws = self.prepare_next_sheet(wb, data_sheet_id, current_page)
                    current_row = self.config['START_ROW']
                    # Re-add validations to the new sheet
                    ws.add_data_validation(dv_g)
                    ws.add_data_validation(dv_q)

                item = all_extracted_data[data_ptr]
                
                # Apply row height
                rd = ws.row_dimensions[current_row]
                rd.height = 20.55
                try: rd.custom_height = True
                except: pass
                
                # [FIX] A~L 범위의 모든 셀에 테두리 적용 (외곽은 굵게)
                for c in range(1, 13):
                    cell = ws.cell(row=current_row, column=c)
                    
                    # 테두리 위치에 따른 스타일 결정
                    l_s = medium_side if c == 1 else thin_side
                    r_s = medium_side if c == 12 else thin_side
                    t_s = medium_side if current_row == self.config['START_ROW'] else thin_side
                    b_s = medium_side if current_row == self.config['DATA_END_ROW'] else thin_side
                    
                    cell.border = Border(left=l_s, right=r_s, top=t_s, bottom=b_s)

                # Drawing No (A~D: 1~4)
                cell_a = ws.cell(row=current_row, column=1)
                cell_a.value = item.get('Dwg', '')
                cell_a.alignment = Alignment(horizontal='center', vertical='center')
                try:
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                except: pass

                # Joint No (F: 6)
                cell_f = ws.cell(row=current_row, column=6)
                cell_f.value = item['No']
                cell_f.alignment = Alignment(horizontal='center', vertical='center')
                # if dv_g: dv_g.add(cell_f) # G열 기준 validation이라 일단 제외하거나 수정 필요

                # Values (Ni: 8, Cr: 9, Mo: 10)
                for val_key, col_idx in [('Ni', 8), ('Cr', 9), ('Mo', 10)]:
                    v = item.get(val_key, 0.0)
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = v if v > 0 else ""; cell.alignment = Alignment(horizontal='center')
                    if v > 0: cell.number_format = '0.0'
                
                # Grade (L: 12)
                cell_l = ws.cell(row=current_row, column=12)
                cell_l.value = str(item.get('Grade', ''))
                cell_l.alignment = Alignment(horizontal='center', vertical='center')
                cell_l.font = Font(size=8.5)
                if dv_q: dv_q.add(cell_l)

                current_row += 1
                data_ptr += 1
                self.progress['value'] = 30 + (data_ptr / len(all_extracted_data)) * 65

            # Footers (Apply page numbers ONLY to Eulji sheets, skipping Gapji)
            total_p = len(wb.worksheets)
            for p_idx, s in enumerate(wb.worksheets):
                if p_idx == 0: continue # Skip the first sheet (Gapji)
                cell_id = "K3"
                try: 
                    s[cell_id] = f"Page {p_idx} of {total_p - 1}" # Exclude Gapji from count if needed, or total sheets
                    s[cell_id].alignment = Alignment(horizontal='distributed', vertical='center')
                except: pass

            # Footer logic handled above
            now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_name = f"{os.path.splitext(os.path.basename(template_path))[0]}_완료_{now_str}{os.path.splitext(template_path)[1]}"
            save_path = os.path.join(os.path.dirname(template_path), output_name)
            wb.save(save_path)
            
            self.progress['value'] = 100
            self.log(f"✨ 완료! 저장됨: {output_name}")
            messagebox.showinfo("성공", f"성적서 생성이 완료되었습니다.\n\n경로: {os.path.dirname(save_path)}\n파일명: {output_name}")
            
        except Exception as e:
            self.log(f"❌ 오류: {e}")
            traceback.print_exc()
            messagebox.showerror("오류", f"작업 중 오류가 발생했습니다:\n{e}")
        finally:
            temp_dir = tempfile.gettempdir()
            for f in glob.glob(os.path.join(temp_dir, "temp_*.png")):
                try: os.remove(f)
                except: pass

if __name__ == "__main__":
    root = tk.Tk()
    PMIReportApp(root)
    root.mainloop()