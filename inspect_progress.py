import json,sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

p=sys.argv[1]; wf=load_workbook(p,data_only=False,keep_links=True);wv=load_workbook(p,data_only=True,keep_links=True)
targets=['갑지(업체제출용)','갑지(롯데내부결재용)','기성내역서(롯데)','실시공 기성내역서(롯데)','1.1 공사건별 수량(케이엔솔and정호이엔씨)']
out={'sheets':{},'external_links':[]}
for x in wf._external_links:
    out['external_links'].append(getattr(x.file_link,'Target',None) or getattr(x.file_link,'target',None))
for sn in targets:
    ws=wf[sn];vs=wv[sn];rows=[]
    for r in range(1,ws.max_row+1):
        cells=[]
        for c in range(1,ws.max_column+1):
            f=ws.cell(r,c).value;v=vs.cell(r,c).value
            if f is not None or v is not None:cells.append({'cell':f'{get_column_letter(c)}{r}','f':f,'v':v})
        if cells:rows.append({'row':r,'cells':cells})
    out['sheets'][sn]=rows
Path(sys.argv[2]).write_text(json.dumps(out,ensure_ascii=False,indent=2,default=str),encoding='utf-8')
