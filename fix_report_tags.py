import openpyxl

file_path = r"C:\Users\jjch2\Desktop\누적진도보고서_202708.xlsx"
wb = openpyxl.load_workbook(file_path)

changed = False
tag_count = 0
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "GS네오텍" in cell.value:
                if cell.column == 2: # B열 (업체)
                    tag_count += 1
                    # 첫번째는 검사현황, 두번째는 불량율 현황
                    tag_str = "[[NDT_121_PAUT]]" if tag_count == 1 else "[[NDT_RESULT_PAUT]]"
                    print(f"Found 'GS네오텍' at {cell.coordinate} in {sheet.title}. Replacing with {tag_str}...")
                    cell.value = tag_str
                    
                    # 같은 행의 다른 데이터(기존 가짜 데이터) 지우기
                    for c in range(3, 22):
                        cell_to_clear = sheet.cell(row=cell.row, column=c)
                        if not isinstance(cell_to_clear, openpyxl.cell.cell.MergedCell):
                            cell_to_clear.value = None
                    changed = True

if changed:
    wb.save(file_path)
    print(f"✅ {file_path} 파일에 태그를 성공적으로 삽입했습니다!")
else:
    print("Could not find dummy GS네오텍 row in the report.")
