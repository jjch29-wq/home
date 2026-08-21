import csv, json, re, sys
from collections import Counter, defaultdict
from pathlib import Path
from openpyxl import load_workbook

def norm(v):
    if v is None: return ''
    if isinstance(v,float) and v.is_integer(): v=int(v)
    return re.sub(r'\s+','',str(v)).upper()

def norm_joint(v):
    s=norm(v)
    m=re.fullmatch(r'0*(\d+)(.*)',s)
    return (str(int(m.group(1)))+m.group(2)) if m else s

def num(v):
    if isinstance(v,(int,float)) and not isinstance(v,bool): return float(v)
    try: return float(str(v).replace(',','').strip())
    except: return 0.0

def extract(path,sheet):
    w=load_workbook(path,data_only=True,read_only=False)
    ws=w[sheet]; rows=[]
    for r in range(5,ws.max_row+1):
        rep=norm(ws.cell(r,2).value)
        if not rep or '-RT-' not in rep: continue
        rec={
          'row':r,'report':rep,'sno':norm(ws.cell(r,3).value),
          'ident':norm(ws.cell(r,4).value),'joint':norm_joint(ws.cell(r,7).value),
          'size':norm(ws.cell(r,8).value),'thk':norm(ws.cell(r,9).value),
          'matl':norm(ws.cell(r,10).value),'welder':norm(ws.cell(r,11).value),
          'film':num(ws.cell(r,12).value),'acc':num(ws.cell(r,13).value),
          'rej':num(ws.cell(r,14).value),'rs':num(ws.cell(r,15).value),
        }
        rec['key']=(rec['report'],rec['sno'])
        rec['joint_key']=(rec['report'],rec['ident'],rec['joint'])
        rows.append(rec)
    return rows

p1,p2,out_json,out_csv=sys.argv[1:5]
a=extract(p1,'정호이엔씨 RT'); b=extract(p2,'RT List')
ma=defaultdict(list); mb=defaultdict(list)
for x in a: ma[x['key']].append(x)
for x in b: mb[x['key']].append(x)

only_a=[]; only_b=[]; mismatches=[]; matched=0
for k in sorted(set(ma)|set(mb)):
    aa=ma.get(k,[]); bb=mb.get(k,[])
    if not bb:
        only_a += aa; continue
    if not aa:
        only_b += bb; continue
    # compare first records; duplicates separately reported
    x,y=aa[0],bb[0]; matched+=1
    diffs={}
    for f in ['ident','joint','size','thk','matl','welder','film','acc','rej','rs']:
        if x[f]!=y[f]: diffs[f]={'jhc':x[f],'status':y[f]}
    if diffs: mismatches.append({'key':list(k),'jhc_row':x['row'],'status_row':y['row'],'diffs':diffs})

# Missing rows that can be located by identification+joint under a different report/SNo.
b_joint=defaultdict(list)
for x in b: b_joint[x['joint_key']].append(x)
relocated=[]; truly_missing=[]
for x in only_a:
    candidates=b_joint.get(x['joint_key'],[]) if all(x['joint_key']) else []
    if candidates: relocated.append({'jhc':x,'status_candidates':candidates})
    else: truly_missing.append(x)

# Authoritative item match by report + identification + joint (S/No may be reordered).
ja=defaultdict(list); jb=defaultdict(list)
for x in a: ja[x['joint_key']].append(x)
for x in b: jb[x['joint_key']].append(x)
item_only_a=[]; item_only_b=[]; item_mismatch=[]; item_matched=0
for k in sorted(set(ja)|set(jb)):
    aa=ja.get(k,[]); bb=jb.get(k,[])
    if not bb: item_only_a+=aa; continue
    if not aa: item_only_b+=bb; continue
    item_matched+=min(len(aa),len(bb))
    x,y=aa[0],bb[0]; diffs={}
    for f in ['size','thk','welder','film','acc','rej','rs']:
        if x[f]!=y[f]: diffs[f]={'jhc':x[f],'status':y[f]}
    if diffs: item_mismatch.append({'key':list(k),'jhc_row':x['row'],'status_row':y['row'],'jhc_sno':x['sno'],'status_sno':y['sno'],'diffs':diffs})

reports_a=defaultdict(lambda:{'rows':0,'film':0,'acc':0,'rej':0,'rs':0})
reports_b=defaultdict(lambda:{'rows':0,'film':0,'acc':0,'rej':0,'rs':0})
for src,dst in [(a,reports_a),(b,reports_b)]:
    for x in src:
        z=dst[x['report']]; z['rows']+=1
        for f in ['film','acc','rej','rs']: z[f]+=x[f]

report_compare=[]
for rep in sorted(reports_a):
    x=reports_a[rep]; y=reports_b.get(rep,{'rows':0,'film':0,'acc':0,'rej':0,'rs':0})
    report_compare.append({'report':rep,'jhc':x,'status':y,'film_delta':x['film']-y['film'],'row_delta':x['rows']-y['rows']})

payload={
 'summary':{
  'jhc_rows':len(a),'jhc_film':sum(x['film'] for x in a),'jhc_acc':sum(x['acc'] for x in a),'jhc_rej':sum(x['rej'] for x in a),'jhc_rs':sum(x['rs'] for x in a),
  'status_rows':len(b),'status_film':sum(x['film'] for x in b),'status_acc':sum(x['acc'] for x in b),'status_rej':sum(x['rej'] for x in b),'status_rs':sum(x['rs'] for x in b),
  'matched_keys':matched,'mismatch_keys':len(mismatches),'only_jhc':len(only_a),'only_status_all':len(only_b),'relocated':len(relocated),'truly_missing':len(truly_missing),
  'duplicate_keys_jhc':sum(1 for v in ma.values() if len(v)>1),'duplicate_keys_status':sum(1 for v in mb.values() if len(v)>1),
  'item_matched':item_matched,'item_only_jhc':len(item_only_a),'item_only_status_all':len(item_only_b),'item_mismatch':len(item_mismatch),
 },
 'report_compare':report_compare,'mismatches':mismatches,'only_jhc':only_a,'relocated':relocated,'truly_missing':truly_missing,
 'item_only_jhc':item_only_a,'item_only_status_all':item_only_b,'item_mismatch':item_mismatch,
}
Path(out_json).write_text(json.dumps(payload,ensure_ascii=False,indent=2,default=str),encoding='utf-8')
with open(out_csv,'w',newline='',encoding='utf-8-sig') as f:
    w=csv.writer(f); w.writerow(['구분','Report No','S/No','Identification No','Joint No','JHC 행','Status 행','JHC Film','Status Film','차이/비고'])
    for m in mismatches:
        d=m['diffs']; w.writerow(['값 불일치',m['key'][0],m['key'][1],d.get('ident',{}).get('jhc',''),d.get('joint',{}).get('jhc',''),m['jhc_row'],m['status_row'],d.get('film',{}).get('jhc',''),d.get('film',{}).get('status',''),json.dumps(d,ensure_ascii=False)])
    for x in truly_missing: w.writerow(['Status 누락',x['report'],x['sno'],x['ident'],x['joint'],x['row'],'',x['film'],'',''])
    for z in relocated:
        x=z['jhc']; cs=z['status_candidates']; w.writerow(['다른 Report/SNo에서 발견',x['report'],x['sno'],x['ident'],x['joint'],x['row'],'/'.join(str(c['row']) for c in cs),x['film'],'/'.join(str(c['film']) for c in cs),''])
