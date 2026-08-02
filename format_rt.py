import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
if 'RT' in wb.sheetnames:
    del wb['RT']
ws = wb.create_sheet('RT', index=1)

headers = ['순번', '제조사', '구분', '촬영일자', 'Section', 'Line No.', 'Joint', '용접사 번호', '촬영구간', '', '', '', '', '', '', '', '관경', '필름규격', '장수', 'R', '촬영결과', '비고']
ws.append(headers)

fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
font = Font(bold=True)
align_center = Alignment(horizontal='center', vertical='center')
align_center_across = Alignment(horizontal='centerContinuous', vertical='center')

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = fill
    cell.border = border
    cell.font = font
    if 9 <= col_idx <= 16:
        cell.alignment = align_center_across
    else:
        cell.alignment = align_center

widths = [6, 10, 6, 12, 15, 30, 8, 12] + [5]*8 + [10, 15, 6, 6, 10, 15]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

dummy_data = [
    [1363, '세경', '', '24.02.28', 'Sec.6', 'HC B 117 300A 02', 'R15', 'WD03', 1, 1, 1, 1, '', '', '', '', '300A', '3 1/3 x 12"', 4, '', '합격', ''],
    [1364, '세경', '', '24.02.28', 'Sec.6', 'HC B 117 300A 02', 'R16', 'WD03', 1, 1, 1, 1, '', '', '', '', '300A', '3 1/3 x 12"', 4, '', '합격', '']
]
for row_data in dummy_data:
    ws.append(row_data)
    for cell in ws[ws.max_row]:
        cell.border = border
        cell.alignment = align_center

ws.auto_filter.ref = f'A1:V{ws.max_row}'
wb.save(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
