import json, sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def summarize(path):
    wf=load_workbook(path,data_only=False,read_only=False)
    wv=load_workbook(path,data_only=True,read_only=False)
    out={'path':str(path),'sheets':[]}
    for ws in wf.worksheets:
        vs=wv[ws.title]
        non=[]
        for row in ws.iter_rows():
            cells=[]
            for c in row:
                if c.value is not None or vs[c.coordinate].value is not None:
                    cells.append({'cell':c.coordinate,'formula':c.value,'value':vs[c.coordinate].value})
            if cells:
                non.append({'row':row[0].row,'cells':cells})
        out['sheets'].append({'name':ws.title,'state':ws.sheet_state,'max_row':ws.max_row,'max_col':ws.max_column,'rows':non})
    return out

payload={'files':[summarize(Path(x)) for x in sys.argv[1:3]]}
Path(sys.argv[3]).write_text(json.dumps(payload,ensure_ascii=False,indent=2,default=str),encoding='utf-8')
