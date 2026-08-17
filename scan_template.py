import openpyxl

file_path = r"C:\Users\jjch2\Desktop\템플릿_최종완성본_V70.xlsx"
try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "화성지사" in cell.value or "화  성  지  사" in cell.value:
                        print(f"Found '화성지사' in {sheet_name}!{cell.coordinate}: {cell.value}")
                    # Also look for other potential errors (like wrong year, etc. if apparent, though mainly Hwaseong)
except Exception as e:
    print(f"Error: {e}")
