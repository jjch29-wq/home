import openpyxl
from openpyxl.styles import Border, Side

wb_orig = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\중앙지사 paut 리스트_서식포함.xlsx')
source_sheet = wb_orig.active

wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')

if 'RT' in wb.sheetnames:
    wb['RT'].auto_filter.ref = None

methods = ['PAUT', 'PT', 'MT']
for m in methods:
    if m in wb.sheetnames:
        del wb[m]
        
    new_ws = wb.copy_worksheet(source_sheet)
    new_ws.title = m
    
    if m != 'PAUT':
        for r in new_ws.iter_rows():
            for c in r:
                if isinstance(c.value, str):
                    c.value = c.value.replace('PAUT', m).replace('paut', m.lower())
                    
    thin_side = Side(style='thin')
    max_r = new_ws.max_row
    max_c = new_ws.max_column
    
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            cell = new_ws.cell(row=r, column=c)
            cb = cell.border
            
            left = thin_side if c == 1 else cb.left
            right = thin_side if c == max_c else cb.right
            top = thin_side if r == 1 else cb.top
            bottom = thin_side if r == max_r else cb.bottom
            
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

wb.move_sheet(wb['PAUT'], offset=-wb.sheetnames.index('PAUT'))
wb.move_sheet(wb['RT'], offset=1-wb.sheetnames.index('RT'))
wb.move_sheet(wb['PT'], offset=2-wb.sheetnames.index('PT'))
wb.move_sheet(wb['MT'], offset=3-wb.sheetnames.index('MT'))

wb.save(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
print("Final fix applied successfully.")
