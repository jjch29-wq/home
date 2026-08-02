import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')

# 1. Delete existing PAUT sheet
if 'PAUT' in wb.sheetnames:
    del wb['PAUT']

# 2. Copy RT sheet to be the new PAUT sheet
rt_sheet = wb['RT']
paut_sheet = wb.copy_worksheet(rt_sheet)
paut_sheet.title = 'PAUT'

# 3. Move the new PAUT sheet to the first position
# In openpyxl, you can move sheets using move_sheet or reordering _sheets
# Let's find its index and move it
paut_idx = wb.sheetnames.index('PAUT')
# move to index 0
wb.move_sheet(paut_sheet, offset=-paut_idx)

wb.save(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
print("PAUT sheet updated with new format and moved to the front.")
