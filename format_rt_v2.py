import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
if 'RT' in wb.sheetnames:
    del wb['RT']
ws = wb.create_sheet('RT', index=1)

# Set Row 1
row1 = ['순번', '제조사', '구분', '촬영\n일자', 'Section', 'Line No.', 'Joint', '용접사\n번호', 
        '촬영구간', '', '', '', '', '', '', '', '관경', '필름규격', '장수', 'R', '촬영결과', '비고']
ws.append(row1)

# Set Row 2
row2 = ['', '', '', '', '', '', '', '', '1', '2', '3', '4', '5', '6', '7', '8', '', '', '', '', '', '']
ws.append(row2)

# Merge cells
merges = [
    'A1:A2', 'B1:B2', 'C1:C2', 'D1:D2', 'E1:E2', 'F1:F2', 'G1:G2', 'H1:H2',
    'I1:P1',
    'Q1:Q2', 'R1:R2', 'S1:S2', 'T1:T2', 'U1:U2', 'V1:V2'
]
for merge in merges:
    ws.merge_cells(merge)

# Styling for headers
fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
font = Font(bold=True)
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=22):
    for cell in row:
        cell.fill = fill
        cell.border = border
        cell.font = font
        cell.alignment = align_center

# Column widths
widths = [6, 10, 6, 12, 12, 30, 8, 10] + [5]*8 + [10, 15, 6, 6, 10, 15]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Add dummy data
dummy_data = [
    [1, '세경', '', '21.09.11', 'SEC 4', 'HC B 112 300A', 'R01', 'WD04', '1', '1', '1', '1', '', '', '', '', '300A', '3 1/3 x 12"', 4, '', '합격', ''],
    [2, '세경', '', '21.09.11', 'SEC 4', 'HC B 112 300A', 'R02', 'WD04', 'P/1', '1', '1', '1', '', '', '', '', '300A', '3 1/3 x 12"', 4, '', '합격', ''],
    [3, '세경', '', '21.09.11', 'SEC 4', 'HC B 112 300A', 'R03', 'WD04', 'P/1', '1', '1', '1', '', '', '', '', '300A', '3 1/3 x 12"', 4, '', '합격', '']
]
for row_data in dummy_data:
    ws.append(row_data)
    for cell in ws[ws.max_row]:
        cell.border = border
        cell.alignment = align_center

# We can set autofilter on row 2, which is standard for merged headers
ws.auto_filter.ref = f'A2:V{ws.max_row}'

wb.save(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
print("Format V2 applied successfully")
