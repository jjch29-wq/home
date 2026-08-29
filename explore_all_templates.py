import openpyxl
import os

templates_dir = r'c:\Users\jjch2\Desktop\PMI\home\src\templates'
for f in os.listdir(templates_dir):
    if f.endswith('.xlsx') and not f.startswith('~'):
        path = os.path.join(templates_dir, f)
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                for row in range(1, 1500):
                    for col in range(1, 15):
                        val = ws.cell(row=row, column=col).value
                        if val and isinstance(val, str) and ('SEC.16' in val or '사진' in val):
                            print(f"[{f}] [{sheetname}] row {row}, col {col}: {val}")
        except Exception as e:
            print(f"Failed to read {f}: {e}")
