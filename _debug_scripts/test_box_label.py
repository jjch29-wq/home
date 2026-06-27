import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_box_label(df_merged, out_path):
    # df_merged는 Sub-Total 행들이 있는 DataFrame
    # 1. Sub-Total 행들만 추출
    if '_is_total' in df_merged.columns:
        df_sub = df_merged[df_merged['_is_total'] == True].copy()
    else:
        # Fallback
        mask = df_merged.astype(str).apply(lambda row: row.str.contains("Sub-Total|소계", case=False).any(), axis=1)
        df_sub = df_merged[mask].copy()

    # Grand Total 제거
    df_sub = df_sub[~df_sub.astype(str).apply(lambda row: row.str.contains("Grand Total|총합계", case=False).any(), axis=1)]
    
    # 2. 필요한 컬럼 추출
    def get_col(synonyms, default=None):
        return next((c for c in df_sub.columns if str(c).lower().strip() in synonyms or any(s in str(c).lower() for s in synonyms)), default)

    report_col = get_col(['report', '성적서'])
    joint_col = get_col(['joint', '조인트'])
    film_col = get_col(['film', '필름'])
    defect_col = get_col(['defect', '결함', 'repair'])
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Box Label"
    
    # Styles
    font_bold = Font(name='맑은 고딕', bold=True, size=11)
    font_normal = Font(name='맑은 고딕', size=11)
    font_small = Font(name='맑은 고딕', size=9)
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 2
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    row_idx = 1
    box_num = 1
    
    for i in range(0, len(df_sub), 2):
        row1 = df_sub.iloc[i] if i < len(df_sub) else None
        row2 = df_sub.iloc[i+1] if i+1 < len(df_sub) else None
        
        for col_offset, row_data in [(0, row1), (4, row2)]:
            if row_data is None:
                continue
                
            report_val = str(row_data.get(report_col, '')) if report_col else ''
            joint_val = str(row_data.get(joint_col, '')) if joint_col else ''
            film_val = str(row_data.get(film_col, '')) if film_col else ''
            defect_val = str(row_data.get(defect_col, '')) if defect_col else ''
            
            # Row 1: BOX NO
            c_box_label = ws.cell(row=row_idx, column=1 + col_offset, value=f"BOX NO. < {box_num} >")
            c_box_label.font = font_bold
            c_box_label.alignment = align_center
            ws.merge_cells(start_row=row_idx, start_column=1 + col_offset, end_row=row_idx, end_column=3 + col_offset)
            
            # Row 2: Report No
            c_rep_label = ws.cell(row=row_idx+1, column=1 + col_offset, value="Report No.")
            c_rep_label.font = font_bold
            c_rep_label.alignment = align_center
            c_rep_val = ws.cell(row=row_idx+1, column=2 + col_offset, value=report_val)
            c_rep_val.font = font_normal
            c_rep_val.alignment = align_center
            ws.merge_cells(start_row=row_idx+1, start_column=2 + col_offset, end_row=row_idx+1, end_column=3 + col_offset)
            
            # Row 3: JOINT NO
            c_jnt_label = ws.cell(row=row_idx+2, column=1 + col_offset, value="JOINT NO.")
            c_jnt_label.font = font_bold
            c_jnt_label.alignment = align_center
            c_jnt_val = ws.cell(row=row_idx+2, column=2 + col_offset, value=joint_val)
            c_jnt_val.font = font_normal
            c_jnt_val.alignment = align_center
            ws.merge_cells(start_row=row_idx+2, start_column=2 + col_offset, end_row=row_idx+2, end_column=3 + col_offset)
            
            # Row 4: 촬영매수 / ORIGINAL
            c_film_label = ws.cell(row=row_idx+3, column=1 + col_offset, value="촬영매수")
            c_film_label.font = font_bold
            c_film_label.alignment = align_center
            ws.merge_cells(start_row=row_idx+3, start_column=1 + col_offset, end_row=row_idx+4, end_column=1 + col_offset)
            
            c_ori_label = ws.cell(row=row_idx+3, column=2 + col_offset, value="ORIGINAL")
            c_ori_label.font = font_normal
            c_ori_label.alignment = align_center
            c_ori_val = ws.cell(row=row_idx+3, column=3 + col_offset, value=film_val)
            c_ori_val.font = font_normal
            c_ori_val.alignment = align_center
            
            # Row 5: REPAIR
            c_rep_label2 = ws.cell(row=row_idx+4, column=2 + col_offset, value="REPAIR")
            c_rep_label2.font = font_normal
            c_rep_label2.alignment = align_center
            c_rep_val2 = ws.cell(row=row_idx+4, column=3 + col_offset, value=defect_val)
            c_rep_val2.font = font_normal
            c_rep_val2.alignment = align_center
            
            # Row 6: SITCO
            c_sitco = ws.cell(row=row_idx+5, column=1 + col_offset, value="SITCO 검사(인)")
            c_sitco.font = font_bold
            c_sitco.alignment = align_left
            ws.merge_cells(start_row=row_idx+5, start_column=1 + col_offset, end_row=row_idx+5, end_column=3 + col_offset)
            
            # Apply borders
            for r in range(row_idx, row_idx+6):
                for c in range(1 + col_offset, 4 + col_offset):
                    ws.cell(row=r, column=c).border = border_all
                    
            box_num += 1
            
        row_idx += 7  # spacing between labels
        
    wb.save(out_path)
    return out_path
