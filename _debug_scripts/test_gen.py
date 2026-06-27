import pandas as pd
import os

# Create dummy subtotal row
df_sub = pd.DataFrame([{
    '_is_total': True,
    'Report': 'SIT-K1-JHC-PIP-RT-0001',
    'Joint': '30',
    'Film': '90',
    'Defect': '0'
}])

def test_gen(out_path):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    font_bold = Font(name='맑은 고딕', bold=True, size=11)
    font_normal = Font(name='맑은 고딕', size=11)
    align_center = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
    align_left = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)
    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_idx = 1
    box_num = 1
    columns = 2
    report_col = 'Report'
    joint_col = 'Joint'
    film_col = 'Film'
    defect_col = 'Defect'

    for i in range(0, len(df_sub), columns):
        items = []
        for j in range(columns):
            r_data = df_sub.iloc[i+j] if i+j < len(df_sub) else None
            items.append((j, j*4, r_data))
        
        for repeat in range(2):
            for j, col_offset, row_data in items:
                if row_data is None: continue
                    
                report_val = str(row_data.get(report_col, ''))
                joint_val = str(row_data.get(joint_col, ''))
                film_val = str(row_data.get(film_col, ''))
                defect_val = str(row_data.get(defect_col, ''))
                
                current_box = box_num + j
                
                c_box_label = ws.cell(row=row_idx, column=1 + col_offset, value=f"BOX NO. < {current_box} >")
                
                c_rep_label = ws.cell(row=row_idx+1, column=1 + col_offset, value="Report No.")
                c_rep_val = ws.cell(row=row_idx+1, column=2 + col_offset, value=report_val)
                
                if repeat == 0:
                    c_jnt_label = ws.cell(row=row_idx+2, column=1 + col_offset, value="Inspection point.")
                    c_jnt_val = ws.cell(row=row_idx+2, column=2 + col_offset, value=joint_val)
                    
                    c_film_label = ws.cell(row=row_idx+3, column=1 + col_offset, value="필름 매수")
                    c_ori_label = ws.cell(row=row_idx+3, column=2 + col_offset, value="ORIGINAL")
                    c_ori_val = ws.cell(row=row_idx+3, column=3 + col_offset, value=f"{film_val} 매" if str(film_val).strip() else "")
                    
                    c_rep_label2 = ws.cell(row=row_idx+4, column=2 + col_offset, value="REPAIR")
                    c_rep_val2 = ws.cell(row=row_idx+4, column=3 + col_offset, value=f"{defect_val} 매" if str(defect_val).strip() else "")
                else:
                    c_jnt_label = ws.cell(row=row_idx+2, column=1 + col_offset, value="Inspection point.")
                    c_jnt_val = ws.cell(row=row_idx+2, column=2 + col_offset, value=joint_val)
                
                c_sitco = ws.cell(row=row_idx+5, column=1 + col_offset, value="SITCO 서울검사(주)")
                
                row_idx += 7
            
        box_num += columns
        
    wb.save(out_path)
    print("Done")

test_gen("test_output.xlsx")
