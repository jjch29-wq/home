import openpyxl
from openpyxl.styles import Border, Side, Alignment
import copy

wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')

hair_side = Side(style='hair')
thin_side = Side(style='thin')

for sheet_name in ['PAUT', 'PT', 'MT']:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_r = ws.max_row
        
        ws.column_dimensions['J'].width = ws.column_dimensions['I'].width
        
        for r in range(1, max_r + 1):
            cell_I = ws.cell(row=r, column=9)
            cell_J = ws.cell(row=r, column=10)
            
            cell_J.value = cell_I.value
            
            new_J_top = thin_side if r <= 2 else hair_side
            new_J_bottom = thin_side if r == max_r else hair_side
            cell_J.border = Border(left=hair_side, right=thin_side, top=new_J_top, bottom=new_J_bottom)
            cell_J.alignment = Alignment(horizontal='center', vertical='center')
            cell_J.fill = copy.copy(cell_I.fill)
            cell_J.font = copy.copy(cell_I.font)
            
            new_I_top = thin_side if r <= 2 else hair_side
            new_I_bottom = thin_side if r == max_r else hair_side
            cell_I.border = Border(left=hair_side, right=hair_side, top=new_I_top, bottom=new_I_bottom)
            
            if r >= 3:
                cell_I.value = None
            elif r == 1:
                cell_I.value = f"{sheet_name}길이"
            elif r == 2:
                cell_I.value = None

wb.save(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
print("Successfully inserted Length column at I, moved Result to J.")
