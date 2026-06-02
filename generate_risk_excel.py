import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

def create_pt_ut_risk_assessment_form(output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "위험성 평가표"

    # 기본 스타일 정의
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(name='맑은 고딕', size=14, bold=True)
    bold_font = Font(name='맑은 고딕', size=11, bold=True)
    normal_font = Font(name='맑은 고딕', size=10)
    bg_fill = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid')

    # 열 너비 설정 (A: 20, B: 60, C: 30)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 85
    ws.column_dimensions['C'].width = 25

    # 1. 문서 제목
    ws['A1'] = "1. 순회점검에 의한 유해·위험요인 조사"
    ws['A1'].font = Font(name='맑은 고딕', size=12, bold=True)
    ws.merge_cells('A1:C1')

    # 2. 표 제목
    ws['A3'] = "사업장 순회점검에 의한 유해·위험요인 조사표"
    ws['A3'].font = header_font
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A3'].fill = bg_fill
    ws.merge_cells('A3:C3')
    for cell in ws['A3:C3'][0]:
        cell.border = thin_border
    ws.row_dimensions[3].height = 35

    # 3. 실시방법
    ws['A4'] = "실시방법"
    ws['A4'].font = bold_font
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A4'].border = thin_border

    ws['B4'] = "위험성평가 수행자가 정기적으로 사업장을 순회점검하고 이 조사표를 사용하여\n유해·위험요인을 찾음"
    ws['B4'].font = normal_font
    ws['B4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.merge_cells('B4:C4')
    for cell in ws['B4:C4'][0]:
        cell.border = thin_border
    ws.row_dimensions[4].height = 40

    # 4. 수행자, 감독, 일시 정보
    info_text = (
        "수행자 성명 :                     (인)                 (인)                 (인)\n"
        "감 독 성명 :                      (인)\n"
        "수행 일시 :  202X.XX.XX"
    )
    ws['A5'] = info_text
    ws['A5'].font = normal_font
    ws['A5'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.merge_cells('A5:C5')
    for cell in ws['A5:C5'][0]:
        cell.border = thin_border
    ws.row_dimensions[5].height = 50

    # 5. 헤더 (유해위험작업 / 사고, 질병 유형)
    ws['A6'] = "유해·위험작업"
    ws['A6'].font = bold_font
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A6:B6')
    
    ws['C6'] = "사고, 질병의 유형"
    ws['C6'].font = bold_font
    ws['C6'].alignment = Alignment(horizontal='center', vertical='center')

    for col in ['A', 'B', 'C']:
        ws[f'{col}6'].border = thin_border

    # 6. 본문 데이터 (PT/UT 변환 데이터)
    risk_factors = [
        ("1. 작업현장 내 발생된 물기 및 폐기물 등에 의해 넘어짐 위험(중)", "1. ② 넘어짐"),
        ("2. 자재 및 장비 이동시 과도한 무게 취급에 따른 근로자 근골격계 질환 발생 위험(중)", "2. ⑫ 불균형 및 무리한 동작"),
        ("3. [PT] 밀폐된 공간이나 환기가 불충분한 곳에서 세척액 및 침투액 사용 시 유기용제 증기 흡입에 의한 중독 위험(최대)", "3. ② 중독"),
        ("4. [PT] 인화성 에어로졸(세척액, 현상액) 취급 중 주변 용접 불꽃 등 화기 접촉으로 인한 화재 및 폭발 위험(최대)", "4. ⑪ 화재"),
        ("5. [PT] 침투액 등 화학물질 취급 시 보호구 미착용으로 인한 피부 접촉 및 피부염 발생 위험(중)", "5. ⑭ 화학물질 누출·접촉"),
        ("6. [UT] 고소작업(비계, 사다리 등) 중 초음파탐상검사 시 안전대 미체결 및 부주의에 의한 추락 위험(최대)", "6. ① 떨어짐"),
        ("7. [UT] 좁고 불편한 자세로 장시간 탐촉자를 문지르는 반복작업에 의한 근골격계 질환 위험(중)", "7. ⑫ 불균형 및 무리한 동작"),
        ("8. [UT] 접촉매질(Couplant, 겔 등)이 바닥에 흘러 작업자가 밟고 미끄러짐(넘어짐) 위험(중)", "8. ② 넘어짐"),
        ("9. [UT] 탐상장비 전원 케이블 피복 손상 또는 습윤한 환경에서 작업 시 감전 위험(최대)", "9. ⑨ 감전"),
        ("10. 통제되지 않은 타공정과 혼재작업 시 중량물 낙하 등에 의한 맞음 사고 위험(최대)", "10. ⑤ 맞음"),
        ("11. 협소한 밀폐공간 내 작업시 질식 또는 약품 흡입 위험(최대)", "11. ② 중독"),
        ("12. 야간작업 시, 어두운 환경에서 이동 및 작업할 경우 시야 미확보에 의한 넘어짐 및 부딪힘 위험(최대)", "12. ② 넘어짐"),
        ("13. [PT] 폐기물(사용한 걸레, 빈 캔)의 무단 방치 및 화기 근접으로 인한 화재 위험(중)", "13. ⑪ 화재"),
    ]

    row_idx = 7
    for task, type_ in risk_factors:
        ws.merge_cells(f'A{row_idx}:B{row_idx}')
        ws[f'A{row_idx}'] = task
        ws[f'A{row_idx}'].font = normal_font
        ws[f'A{row_idx}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        ws[f'C{row_idx}'] = type_
        ws[f'C{row_idx}'].font = normal_font
        ws[f'C{row_idx}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # 테두리
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row_idx}'].border = thin_border
            
        ws.row_dimensions[row_idx].height = 30
        row_idx += 1

    # 7. 하단 범례 (사고의 유형, 질병의 유형)
    legend_row = row_idx
    ws.merge_cells(f'A{legend_row}:B{legend_row}')
    ws[f'A{legend_row}'] = (
        "[사고의 유형]\n"
        "① 떨어짐   ② 넘어짐   ③ 깔림   ④ 부딪힘   ⑤ 맞음   ⑥ 무너짐   ⑦ 끼임   ⑧ 절단·베임·찔림\n"
        "⑨ 감전   ⑩ 폭발·파열   ⑪ 화재   ⑫ 불균형 및 무리한 동작   ⑬ 이상온도·물체접촉   ⑭ 화학물질 누출·접촉   ⑮ 기타"
    )
    ws[f'A{legend_row}'].font = normal_font
    ws[f'A{legend_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    ws[f'C{legend_row}'] = (
        "[질병의 유형]\n"
        "① 진폐\n"
        "② 중독\n"
        "③ 난청\n"
        "④ 요통\n"
        "⑤ 기타"
    )
    ws[f'C{legend_row}'].font = normal_font
    ws[f'C{legend_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    for col in ['A', 'B', 'C']:
        ws[f'{col}{legend_row}'].border = thin_border
    ws.row_dimensions[legend_row].height = 90

    # 여백 설정
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    wb.save(output_path)
    print(f"엑셀 파일 생성 완료: {output_path}")

if __name__ == '__main__':
    output_excel = r"C:\Users\-\OneDrive\바탕 화면\3. 유해 위험요인 조사표_PT_UT_수정본.xlsx"
    create_pt_ut_risk_assessment_form(output_excel)
