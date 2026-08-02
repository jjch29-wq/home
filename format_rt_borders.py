import openpyxl
from openpyxl.styles import Border, Side

wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')

if 'RT' in wb.sheetnames:
    ws = wb['RT']
    
    # Define dashed border for data rows
    dashed_side = Side(style='dashed')
    thin_side = Side(style='thin')
    
    # We keep thin outer borders, but dashed inner borders. 
    # For simplicity, let's make all borders dashed for data cells, 
    # except the very left/right if we want, but the image shows dashed everywhere inside.
    dashed_border = Border(left=dashed_side, right=dashed_side, top=dashed_side, bottom=dashed_side)
    
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=22):
        for cell in row:
            cell.border = dashed_border

# If user meant applying RT style to PT and MT, we should probably format them too.
# For now, let's just fix RT's borders and ask if they want it applied to others.
wb.save(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
print("Dashed borders applied to RT data rows.")
