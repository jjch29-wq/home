import openpyxl

wb = openpyxl.load_workbook(r'C:/Users/jjch2/Desktop/템플릿_최종완성본_V74.xlsx')
ws = wb.worksheets[0]

# Header rows to KEEP: 403, 404 (PAUT section 1), 448, 449 (PAUT section 2)
# TOTAL rows to KEEP: 481, 482
# Section title rows to KEEP: (anything with section title text)

KEEP_ROWS = {403, 404, 448, 449, 481, 482}

# Values to KEEP (structural labels)
KEEP_VALUES = {'TOTAL', 'ORI', 'RE', "ORI'", "RE'"}

# Known data values to ALWAYS clear
def is_data_value(val):
    if val is None:
        return False
    v = str(val).strip()
    if not v:
        return False
    # These are data values, not structural
    data_indicators = ['GS네오텍', 'Sec.', 'JA2026', '1.2767', '주간', '야간', '합격', '불합격']
    for ind in data_indicators:
        if ind in v:
            return True
    # Numbers that are clearly data (not row labels)
    try:
        f = float(v)
        if f > 0:
            return True
    except:
        pass
    # Single letter 'M' in certain columns is data
    if v == 'M':
        return True
    return False

cleared = 0
# Clear ALL data from rows 405~480 (data area before TOTAL)
for row in range(405, 481):
    if row in KEEP_ROWS:
        continue
    for col in range(1, 30):
        cell = ws.cell(row=row, column=col)
        if is_data_value(cell.value):
            cell.value = None
            cleared += 1

print(f"Cleared {cleared} data cells")

# Verify what remains
print("\n=== Remaining non-empty rows (405-480) ===")
for row in range(405, 481):
    cells = {}
    for col in range(1, 30):
        c = ws.cell(row=row, column=col)
        if c.value is not None:
            v = str(c.value).strip()
            if v:
                cells[col] = v
    if cells:
        print(f"  Row {row}: {cells}")

wb.save(r'C:/Users/jjch2/Desktop/템플릿_최종완성본_V74.xlsx')
print("\nSaved!")
