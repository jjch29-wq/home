import codecs

file_path = r'c:\Users\jjch2\Desktop\PMI\home\src\한국지역난방 중앙지사.py'
with codecs.open(file_path, 'r', 'utf-8') as f:
    text = f.read()

target1 = """                if not filepath:
                    messagebox.showwarning("입력 오류", "템플릿 파일을 선택해주세요.")
                    return"""
replacement1 = """                if not filepath:
                    messagebox.showwarning("입력 오류", "템플릿 파일을 선택해주세요.")
                    return
                
                save_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    initialfile=f"월간진도보고서_{year}년_{month:02d}월.xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    title="월간진도보고서 통합 저장"
                )
                if not save_path:
                    return"""

text = text.replace(target1, replacement1)

target2 = """                # --- 5. 엑셀 기입 ---
                import openpyxl
                wb = openpyxl.load_workbook(filepath)"""
replacement2 = """                # --- 5. 엑셀 기입 (통합 처리) ---
                try:
                    from monthly_report_exporter import MonthlyReportExporter
                    import json, os
                    history = {}
                    history_path = os.path.join(self.data_dir if hasattr(self, 'data_dir') else 'data', "daily_work_history.json")
                    if os.path.exists(history_path):
                        with open(history_path, 'r', encoding='utf-8') as f:
                            history = json.load(f)
                    
                    target_month_str = f"{year}-{month:02d}"
                    exporter = MonthlyReportExporter(history, target_month_str, filepath, doc_num)
                    exporter.generate(save_path)
                except Exception as e:
                    messagebox.showerror("오류", f"태그 변환 중 오류: {e}")
                    import traceback
                    log(traceback.format_exc())
                    return

                import openpyxl
                wb = openpyxl.load_workbook(save_path)"""

text = text.replace(target2, replacement2)

target3 = """                    ws.cell(row=start_row, column=2, value="전체 총계(TOTAL)")
                    ws.cell(row=start_row, column=17, value=round(tot_ori, 4) if tot_ori else '-')
                    ws.cell(row=start_row, column=18, value=round(tot_re, 4) if tot_re else '-')
                    ws.cell(row=start_row, column=20, value=round(tot_total, 4) if tot_total else '-')
                    
                except Exception as e:"""
replacement3 = """                    ws.cell(row=start_row, column=2, value="전체 총계(TOTAL)")
                    ws.cell(row=start_row, column=17, value=round(tot_ori, 4) if tot_ori else '-')
                    ws.cell(row=start_row, column=18, value=round(tot_re, 4) if tot_re else '-')
                    ws.cell(row=start_row, column=20, value=round(tot_total, 4) if tot_total else '-')
                    
                wb.save(save_path)
                messagebox.showinfo("완료", "월간 진도보고서 작성이 모두 완료되었습니다!")
                os.startfile(os.path.dirname(save_path))
                except Exception as e:"""

text = text.replace(target3, replacement3)

with codecs.open(file_path, 'w', 'utf-8') as f:
    f.write(text)
