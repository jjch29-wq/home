import openpyxl

file_path = r"C:\Users\jjch2\Desktop\템플릿_최종완성본_V70_수정.xlsx"
wb = openpyxl.load_workbook(file_path)

changed = False
tag_count = 0
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "GS네오텍" in cell.value:
                if cell.column == 2: # B열 (업체)
                    tag_count += 1
                    # 1번째는 검사현황, 2번째는 불량율 현황으로 간주
                    tag_str = "[[NDT_121_PAUT]]" if tag_count == 1 else "[[NDT_RESULT_PAUT]]"
                    print(f"Found 'GS네오텍' at {cell.coordinate} in {sheet.title}. Replacing with {tag_str}...")
                    cell.value = tag_str
                    
                    # 같은 행의 다른 데이터 지우기
                    for c in range(3, 22):
                        sheet.cell(row=cell.row, column=c).value = None
                    changed = True

if changed:
    wb.save(file_path)
    print("Template updated with tags.")
else:
    print("Could not find dummy GS네오텍 row in template.")
