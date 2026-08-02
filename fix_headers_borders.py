import openpyxl
from openpyxl.styles import Border, Side

wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')

thin_side = Side(style='thin')
hair_side = Side(style='hair')

for sheet_name in ['PAUT', 'PT', 'MT']:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_r = ws.max_row
        max_c = ws.max_column
        
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                
                # Determine border styles
                left_style = thin_side if (c == 1 or r == 1) else hair_side
                right_style = thin_side if (c == max_c or r == 1) else hair_side
                top_style = thin_side if (r == 1 or r == 2) else hair_side
                bottom_style = thin_side if (r == max_r or r == 1) else hair_side
                
                cell.border = Border(left=left_style, right=right_style, top=top_style, bottom=bottom_style)

wb.save(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
print("Fixed PAUT, PT, MT header and data borders.")
