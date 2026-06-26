import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

def insert_image_to_excel(ws, paths_str, cell_str, cell_h_px, cell_w_px):
    if not paths_str: return
    try:
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from PIL import Image as PILImage
        from PIL import ImageOps
        import os, uuid
        
        paths = [p.strip() for p in paths_str.split("|") if p.strip() and os.path.exists(p.strip())]
        if not paths: return
        
        # 95% of the cell size
        target_h_px = int(cell_h_px * 0.95)
        target_w_px = int(cell_w_px * 0.95)
        
        out_dir = os.path.join(os.path.expanduser("~"), ".gemini", "scratch")
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, f"temp_excel_img_{uuid.uuid4().hex[:8]}.png")

        if len(paths) > 1:
            images = [PILImage.open(p).convert('RGBA') for p in paths]
            min_height = min(img.height for img in images)
            resized_images = [img.resize((int(img.width * min_height / img.height), min_height)) for img in images]
            spacing = 10
            total_width = sum(img.width for img in resized_images) + spacing * (len(resized_images) - 1)
            combined = PILImage.new('RGBA', (total_width, min_height), (255, 255, 255, 0))
            x_offset = 0
            for img in resized_images:
                combined.paste(img, (x_offset, 0))
                x_offset += img.width + spacing
            
            # Center crop to target size
            cropped_img = ImageOps.fit(combined, (target_w_px, target_h_px), method=PILImage.Resampling.LANCZOS)
        else:
            img = PILImage.open(paths[0]).convert('RGBA')
            # Center crop to target size
            cropped_img = ImageOps.fit(img, (target_w_px, target_h_px), method=PILImage.Resampling.LANCZOS)
            
        # Place cropped image into a full cell-sized transparent canvas to act as a margin
        final_img = PILImage.new('RGBA', (cell_w_px, cell_h_px), (255, 255, 255, 0))
        offset_x = (cell_w_px - target_w_px) // 2
        offset_y = (cell_h_px - target_h_px) // 2
        final_img.paste(cropped_img, (offset_x, offset_y))
        
        final_img.save(out_path, format="PNG")
            
        xl_img = OpenpyxlImage(out_path)
        xl_img.width = cell_w_px
        xl_img.height = cell_h_px
        
        ws.add_image(xl_img, cell_str)
    except Exception as e:
        print(f"이미지 삽입 오류: {e}")

def generate_excel(data, output_path):
    wb = openpyxl.Workbook()

    # 첫 번째 시트: 회의자료 본문 (표지와 통합)
    ws = wb.active
    ws.title = "안전보건협의체 회의자료"

    font_title    = Font(name='맑은 고딕', size=16, bold=True)
    font_subtitle = Font(name='맑은 고딕', size=12, bold=True)
    font_bold     = Font(name='맑은 고딕', size=11, bold=True)
    font_normal   = Font(name='맑은 고딕', size=11)
    align_center  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin_border   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30

    def apply_border(range_str):
        for row in ws[range_str]:
            for cell in row:
                cell.border = thin_border

    font_cover_sub = Font(name='맑은 고딕', size=16, bold=True, italic=True, color='4472C4')
    font_cover_title = Font(name='맑은 고딕', size=28, bold=True)

    ws.merge_cells('A1:D1')
    ws['A1'] = "「가산~가평 천연가스 공급시설 건설공사」"
    ws['A1'].font = font_cover_sub
    ws['A1'].alignment = align_center

    ws.merge_cells('A2:D2')
    ws['A2'] = "{} 안전 및 보건에 관한 협의체 회의자료".format(data.get('제출년월', '2026년 7월'))
    ws['A2'].font = font_cover_title
    ws['A2'].alignment = align_center

    apply_border('A5:D6')
    ws.merge_cells('A5:B5'); ws['A5'] = "수급업체명"
    ws['A5'].font = font_bold; ws['A5'].alignment = align_center; ws['A5'].fill = fill_header
    ws.merge_cells('C5:D5'); ws['C5'] = data.get("계약상대자(업체명)", "서울검사(주)")
    ws['C5'].font = font_normal; ws['C5'].alignment = align_center
    ws.merge_cells('A6:B6'); ws['A6'] = "제 출 년 월"
    ws['A6'].font = font_bold; ws['A6'].alignment = align_center; ws['A6'].fill = fill_header
    ws.merge_cells('C6:D6'); ws['C6'] = data.get("제출년월", "2026년 7월")
    ws['C6'].font = font_normal; ws['C6'].alignment = align_center

    # 1~6행이 A4 한 페이지(표지)를 꽉 채우도록 높이 대폭 확대
    ws.row_dimensions[1].height = 200
    ws.row_dimensions[2].height = 150
    ws.row_dimensions[3].height = 120
    ws.row_dimensions[4].height = 120
    ws.row_dimensions[5].height = 40
    ws.row_dimensions[6].height = 40
    ws.row_dimensions[7].height = 30 # 다음 페이지와의 간격

    # 7행 아래에 페이지 나누기 추가 (A4 1페이지로 분리)
    from openpyxl.worksheet.pagebreak import Break
    ws.row_breaks.append(Break(id=7))

    ws.merge_cells('A8:D8')
    ws['A8'] = "1. 수급업체 현황"
    ws['A8'].font = font_subtitle
    ws['A8'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[8].height = 40

    table_data = [
        ("계 약 명",   data.get("계약명", "")),
        ("계약기간",    data.get("계약기간", "")),
        ("계약상대자(업체명)", data.get("계약상대자(업체명)", "")),
        ("현장대리인",  data.get("현장대리인", "")),
        ("작업의 시작시간", data.get("작업의 시작시간", "")),
        ("작업 또는 작업장 간의 연락방법", data.get("작업 또는 작업장 간의 연락방법", "")),
        ("재해발생 위험시의 대피방법", data.get("재해발생 위험시의 대피방법", "")),
        ("사업자와 수급인 또는 수급인 상호간의 연락방법", data.get("사업자와 수급인 또는 수급인 상호간의 연락방법", "현장소장 및 안전관리자 핫라인 구축")),
        ("작업공정의 조정 및 협의 요청사항", data.get("작업공정의 조정 및 협의 요청사항", "")),
        ("{}\n주요 활동실적\n(작업사항)".format(data.get('실적년월', '2026년 6월')), data.get("주요 활동실적", "")),
        ("{}\n주요 활동계획\n(작업사항)".format(data.get('계획년월', '2026년 7월')), data.get("주요 활동계획", "")),
    ]

    start_row = 9
    for i, (label, val) in enumerate(table_data):
        row = start_row + i
        apply_border('A{}:D{}'.format(row, row))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value=label).font = font_bold
        ws.cell(row=row, column=1).alignment = align_center
        ws.cell(row=row, column=1).fill = fill_header
        ws.cell(row=row, column=3, value=val).font = font_normal
        ws.cell(row=row, column=3).alignment = align_left
        if '\n' in label or '\n' in val:
            lines = max(label.count('\n'), val.count('\n')) + 1
            ws.row_dimensions[row].height = lines * 18 + 20
        else:
            ws.row_dimensions[row].height = 35

    current_row = start_row + len(table_data) + 1
    ws.row_dimensions[current_row-1].height = 30
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=1, value="2. 위험성평가 실시 현황").font = font_subtitle
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    headers = ["구분", "실시여부(O,X)", "작성 날짜", "비고"]
    apply_border('A{}:D{}'.format(current_row, current_row))
    for col, h in enumerate(headers, 1):
        ws.cell(row=current_row, column=col, value=h).font = font_bold
        ws.cell(row=current_row, column=col).alignment = align_center
        ws.cell(row=current_row, column=col).fill = fill_header
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    risk_data = [
        ("최초 위험성평가", data.get("최초위험성평가_실시여부", ""),
         data.get("최초위험성평가_작성날짜", ""), "계약당시 전체 작업 대상 작성"),
        ("정기 위험성평가", data.get("정기위험성평가_실시여부", ""),
         data.get("정기위험성평가_작성날짜", ""), "매년 당해연도 전체 작업 대상"),
        ("수시위험성평가",  data.get("수시위험성평가_실시여부", "O"),
         data.get("수시위험성평가_작성날짜", ""), "매월 협의체 회의시 제출"),
    ]
    for row_data in risk_data:
        apply_border('A{}:D{}'.format(current_row, current_row))
        for col, val in enumerate(row_data, 1):
            ws.cell(row=current_row, column=col, value=val).font = font_normal
            ws.cell(row=current_row, column=col).alignment = align_center
            if col == 1:
                ws.cell(row=current_row, column=col).fill = fill_header
                ws.cell(row=current_row, column=col).font = font_bold
            elif col == 4:
                ws.cell(row=current_row, column=col).alignment = align_left
        ws.row_dimensions[current_row].height = 40
        current_row += 1

    # 빠진 부분 추가: 위험성평가서 별첨 및 변동 없음 체크
    apply_border(f'A{current_row}:D{current_row}')
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=1, value="  ☐ 위험성평가서 별첨          ☐ 정기평가와 변동 없음(아래 1항~5항 해당없을시 생략가능)").font = font_normal
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # 빠진 부분 추가: 안내 문구 라인
    apply_border(f'A{current_row}:D{current_row}')
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    info_text = (
        "1. 사업장 건설물의 설치·이전·변경 또는 해체          2. 기계,기구,설비,원재료 등의 신규 도입 또는 변경\n"
        "3. 건설물,기계·기구,설비 등의 정비 또는 보수 (주기적, 반복적 작업으로서 정기평가를 실시한 경우에는 제외)\n"
        "4. 작업방법 또는 작업절차의 신규도입 또는 변경  5. 중대산업사고 또는 산업재해 발생\n"
        "6. 위험요인 추가발굴 및 아차사고보고서 내용 반영"
    )
    ws.cell(row=current_row, column=1, value=info_text).font = Font(name='맑은 고딕', size=9)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[current_row].height = 65
    current_row += 1

    # 25행 뒤에 페이지 나누기를 추가하여 2페이지를 여기서 마감
    from openpyxl.worksheet.pagebreak import Break
    ws.row_breaks.append(Break(id=current_row - 1))

    # --- 3. 위험성평가 중점관리항목 개선사항 ---
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=1, value="3. 위험성평가 중점관리항목 개선사항").font = font_subtitle
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    apply_border(f'A{current_row}:D{current_row+1}')
    ws.cell(row=current_row, column=1, value="관리 주관부서").font = font_bold
    ws.cell(row=current_row, column=1).alignment = align_center
    ws.cell(row=current_row, column=1).fill = fill_header
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=2, value=data.get("관리주관부서", "")).font = font_normal
    ws.cell(row=current_row, column=2).alignment = align_left
    ws.row_dimensions[current_row].height = 45
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="장소").font = font_bold
    ws.cell(row=current_row, column=1).alignment = align_center
    ws.cell(row=current_row, column=1).fill = fill_header
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=2, value=data.get("장소", "")).font = font_normal
    ws.cell(row=current_row, column=2).alignment = align_left
    ws.row_dimensions[current_row].height = 45
    current_row += 1

    apply_border(f'A{current_row}:D{current_row}')
    ws.cell(row=current_row, column=1, value="위험성평가 중점관리항목").font = font_bold
    ws.cell(row=current_row, column=1).alignment = align_center
    ws.cell(row=current_row, column=1).fill = fill_header
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=2, value=data.get("위험성평가_중점관리항목", "")).font = font_normal
    ws.cell(row=current_row, column=2).alignment = align_left
    ws.row_dimensions[current_row].height = 140
    current_row += 1

    apply_border(f'A{current_row}:D{current_row}')
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    ws.cell(row=current_row, column=1, value="위험성 감소대책 수립 및 실행방법 Check").font = font_bold
    ws.cell(row=current_row, column=1).alignment = align_center
    ws.cell(row=current_row, column=1).fill = fill_header
    
    check_texts = []
    if data.get("감소대책_위험성제거"): check_texts.append("1.위험성제거(✔)")
    else: check_texts.append("1.위험성제거( )")
    if data.get("감소대책_공학적"): check_texts.append("2.공학적(✔)")
    else: check_texts.append("2.공학적( )")
    if data.get("감소대책_관리적"): check_texts.append("3.관리적(✔)")
    else: check_texts.append("3.관리적( )")
    if data.get("감소대책_개인보호구"): check_texts.append("4.개인보호구(✔)")
    else: check_texts.append("4.개인보호구( )")
    
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=3, value=" ".join(check_texts)).font = font_normal
    ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
    ws.row_dimensions[current_row].height = 45
    current_row += 1

    apply_border(f'A{current_row}:D{current_row}')
    ws.cell(row=current_row, column=1, value="안전·보건 개선조치 이행사항").font = font_bold
    ws.cell(row=current_row, column=1).alignment = align_center
    ws.cell(row=current_row, column=1).fill = fill_header
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=2, value=data.get("개선조치_이행사항", "")).font = font_normal
    ws.cell(row=current_row, column=2).alignment = align_left
    ws.row_dimensions[current_row].height = 140
    current_row += 1

    # Photos for Section 3
    apply_border(f'A{current_row}:D{current_row+1}')
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    ws.cell(row=current_row, column=1, value="조치 전 사진").font = font_bold
    ws.cell(row=current_row, column=1).alignment = align_center
    ws.cell(row=current_row, column=1).fill = fill_header
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=3, value="조치 후 사진").font = font_bold
    ws.cell(row=current_row, column=3).alignment = align_center
    ws.cell(row=current_row, column=3).fill = fill_header
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
    ws.row_dimensions[current_row].height = 360
    insert_image_to_excel(ws, data.get("조치사진_경로", ""), f"A{current_row}", 470, 384)
    insert_image_to_excel(ws, data.get("개선후사진_경로", ""), f"C{current_row}", 470, 384)
    current_row += 1

    # 3페이지 마감 (Section 3 끝)
    ws.row_breaks.append(Break(id=current_row - 1))

    # --- 4. 아차사고 보고서 ---
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=1, value="4. 아차사고 보고서").font = font_subtitle
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    acha_data = [
        ("사 고 명", data.get("아차사고_사고명", "")),
        ("발생일시", data.get("아차사고_발생일시", "")),
        ("장 소 (설 비)", data.get("아차사고_장소", "")),
        ("보 고 자", data.get("아차사고_보고자", "")),
        ("소    속", data.get("아차사고_소속", "")),
        ("사고내용\n(6하원칙)", data.get("아차사고_사고내용", "")),
        ("문제점 및\n원인분석", data.get("아차사고_원인분석", ""))
    ]
    for label, val in acha_data:
        apply_border(f'A{current_row}:D{current_row}')
        ws.cell(row=current_row, column=1, value=label).font = font_bold
        ws.cell(row=current_row, column=1).alignment = align_center
        ws.cell(row=current_row, column=1).fill = fill_header
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        ws.cell(row=current_row, column=2, value=val).font = font_normal
        ws.cell(row=current_row, column=2).alignment = align_left
        if '\n' in label or '\n' in val:
            ws.row_dimensions[current_row].height = 70
        else:
            ws.row_dimensions[current_row].height = 45
        current_row += 1

    apply_border(f'A{current_row}:D{current_row+1}')
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=1, value="사진대지").font = font_bold
    ws.cell(row=current_row, column=1).alignment = align_center
    ws.cell(row=current_row, column=1).fill = fill_header
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
    ws.row_dimensions[current_row].height = 380
    insert_image_to_excel(ws, data.get("아차사고_조치전사진", ""), f"A{current_row}", 490, 384)
    insert_image_to_excel(ws, data.get("아차사고_조치후사진", ""), f"C{current_row}", 490, 384)
    current_row += 1

    # 4페이지 마감 (Section 4 끝)
    ws.row_breaks.append(Break(id=current_row - 1))

    # --- 5. 안전·보건관련 건의 및 제의사항 ---
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=1, value="5. 안전·보건관련 건의 및 제의사항").font = font_subtitle
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[current_row].height = 50
    current_row += 1

    align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    apply_border(f'A{current_row}:D{current_row+1}')
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=1, value=f"\n  󰏚 개진사항\n    ❍ {data.get('건의_개진사항', '')}").font = font_normal
    ws.cell(row=current_row, column=1).alignment = align_top_left
    ws.row_dimensions[current_row].height = 390
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=1, value=f"\n  󰏚 제안사유\n    ❍ {data.get('건의_제안사유', '')}").font = font_normal
    ws.cell(row=current_row, column=1).alignment = align_top_left
    ws.row_dimensions[current_row].height = 390
    current_row += 1

    # 인쇄 영역 및 페이지 설정 (너비는 1페이지 맞춤, 길이는 자동)
    ws.print_area = f'A1:D{current_row}'
    ws.page_setup.paperSize = 9  # A4 사이즈
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # 세 번째 시트: 교육이수관리대장
    generate_edu_sheet(wb, data)
    wb.save(output_path)
    return output_path


def generate_cover_sheet(ws, data):
    """표지 시트 생성 - 이미지 양식 재현
    열 구성: A(좌여백) B(레이블) C(값) D(우여백)
    """
    yr_month = data.get('제출년월', '2026년 7월')
    company  = data.get('계약상대자(업체명)', '서울검사(주)')

    # 열 너비 설정 (A4 세로 비율에 맞게 확장)
    ws.column_dimensions['A'].width = 4    # 좌 여백
    ws.column_dimensions['B'].width = 26   # 레이블
    ws.column_dimensions['C'].width = 50   # 값 / 제목 영역
    ws.column_dimensions['D'].width = 4    # 우 여백

    # 기본 행 높이 (A4 세로에 꽉 차도록)
    for r in range(1, 44):
        ws.row_dimensions[r].height = 19.5

    blue_line = Side(style='medium', color='1F4E79')
    thin      = Side(style='thin')
    bdr_thin  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(r, c, val='', font=None, align=None, border=None, fill=None):
        cl = ws.cell(row=r, column=c, value=val)
        if font:   cl.font      = font
        if align:  cl.alignment = align
        if border: cl.border    = border
        if fill:   cl.fill      = fill
        return cl

    al = Alignment(horizontal='left',   vertical='center')
    ac = Alignment(horizontal='center', vertical='center')

    fill_hdr = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # ── 행 7: 부제목 (가산~가평...)
    ws.row_dimensions[7].height = 28
    ws.merge_cells('B7:C7')
    c7 = cell(7, 2,
        val='「가산~가평 천연가스 공급시설 건설공사」',
        font=Font(name='맑은 고딕', size=13, italic=True, color='4472C4'),
        align=al)
    c7.border = Border(bottom=blue_line)

    # ── 행 8: 파란 줄 여분 높이
    ws.row_dimensions[8].height = 6
    ws.row_dimensions[9].height = 6

    # ── 행 10: 년월
    ws.row_dimensions[10].height = 36
    ws.merge_cells('B10:C10')
    cell(10, 2,
         val=yr_month,
         font=Font(name='맑은 고딕', size=20, bold=True),
         align=al)

    # ── 행 11~12: 메인 타이틀
    ws.row_dimensions[11].height = 70
    ws.row_dimensions[12].height = 8
    ws.merge_cells('B11:C11')
    c11 = cell(11, 2,
        val='안전  및  보건에  관한  협의체  회의자료',
        font=Font(name='맑은 고딕', size=32, bold=True),
        align=al)
    c11.border = Border(bottom=blue_line)

    # ── 행 36~37: 하단 표 (수급업체명 / 제출년월)
    tbl_data = [
        ('수급업체명',   company),
        ('제  출  년  월', yr_month),
    ]
    for i, (label, val) in enumerate(tbl_data):
        r = 36 + i
        ws.row_dimensions[r].height = 34
        # 레이블
        lc = cell(r, 2, label,
                  font=Font(name='맑은 고딕', size=13, bold=True),
                  align=ac, border=bdr_thin, fill=fill_hdr)
        # 값
        vc = cell(r, 3, val,
                  font=Font(name='맑은 고딕', size=13),
                  align=ac, border=bdr_thin)

    # 인쇄 설정 (A4 세로)
    ws.print_area = 'A1:D40'
    ws.page_setup.paperSize   = 9
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1






# ═══════════════════════════════════════════════════════════════════════
# 시트2: 수급업체 안전보건교육 이수 관리대장 (이미지 양식 구조)
#
# 열 구성 (30열, A~AD):
#   A(1)       : NO
#   B(2)       : 성명(상행) / 생년월일(하행) - 1인 2행
#   C(3)       : 작업(근로)자 Check
#   D(4)       : 관리감독자 상용 Check
#   E(5)       : 관리감독자 임용 Check
#   F(6)       : 채용일자(최초투입일)
#   G(7)       : 채용시교육 일용(1시간)
#   H(8)       : 채용시교육 일용외(8시간)
#   I(9)       : 특별교육 대상여부 - 대상
#   J(10)      : 특별교육 대상여부 - 비대상(작업명)
#   K(11)      : 특별교육 교육이수일
#   L~N(12~14) : 1분기 1,2,3월
#   O~Q(15~17) : 2분기 4,5,6월
#   R~T(18~20) : 3분기 7,8,9월
#   U~W(21~23) : 4분기 10,11,12월
#   X(24)      : 관리감독자 연간16h 매반
#   Y(25)      : 관리감독자 연간16h 3월
#   Z(26)      : 작업자 매반기12h 매반
#   AA(27)     : 작업자 매반기12h 6월
#   AB(28)     : 작업자 매반기12h 매반(하반기)
#   AC(29)     : 작업자 매반기12h 12월
#   AD(30)     : 비고
# ═══════════════════════════════════════════════════════════════════════
def generate_edu_sheet(wb, data):
    ws = wb.create_sheet("안전보건교육 이수 관리대장")

    ft_title  = Font(name='맑은 고딕', size=18, bold=True)
    ft_sub    = Font(name='맑은 고딕', size=14, bold=True)
    ft_bold   = Font(name='맑은 고딕', size=12, bold=True)
    ft_normal = Font(name='맑은 고딕', size=12)
    ft_small  = Font(name='맑은 고딕', size=11)

    ac = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True)
    al = Alignment(horizontal='left',   vertical='center', wrap_text=True, shrink_to_fit=True)
    thin  = Side(style='thin')
    bdr_t = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_hdr  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_spec = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    fill_q1   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_q2   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_q3   = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_q4   = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_mgr  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_wkr  = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_white= PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    LAST_COL = 30

    col_w = {
        1:5,  2:11, 3:9,  4:12, 5:12, 6:13, 7:9,  8:10,
        9:9,  10:13,11:18,
        12:6, 13:6, 14:6,
        15:6, 16:6, 17:6,
        18:6, 19:6, 20:6,
        21:6, 22:6, 23:6,
        24:7, 25:7,
        26:7, 27:7, 28:7, 29:7,
        30:18,
    }
    for c, w in col_w.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    def bc(r, c, val="", font=None, fill=None, align=None):
        cell = ws.cell(row=r, column=c, value=val)
        if font:  cell.font  = font
        if fill:  cell.fill  = fill
        if align: cell.alignment = align
        return cell

    def apply_bdr(r1, c1, r2, c2):
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(row=r, column=c).border = bdr_t

    def mg(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    yr_month = data.get("교육관리_년월",      data.get("제출년월", "2026년 6월"))
    company  = data.get("교육관리_수급업체명", data.get("계약상대자(업체명)", ""))
    contract = data.get("교육관리_계약명",     data.get("계약명", ""))
    manager  = data.get("교육관리_현장대리인", data.get("현장대리인", ""))
    records  = data.get("교육관리_직원목록",   [])

    for rn, ht in [(1,50),(2,30),(3,40),(4,60),(5,50),(6,45)]:
        ws.row_dimensions[rn].height = ht

    # ── 행1: 제목
    mg(1,1,1,LAST_COL)
    bc(1,1, "'26년  {}  수급업체 안전보건교육 이수 관리대장".format(yr_month),
       font=ft_title, align=ac)

    # ── 행2: 수급업체명 / 계약명
    mg(2,1,2,2);   bc(2,1, "- 수급업체명 :", font=ft_sub, align=al)
    mg(2,3,2,10);  bc(2,3, company,          font=ft_sub, align=al)
    mg(2,11,2,13); bc(2,11,"- 계  약  명 :", font=ft_sub, align=al)
    mg(2,14,2,LAST_COL); bc(2,14, contract,  font=ft_sub, align=al)

    # ── 행3: 확인 문구
    mg(3,1,3,19)
    bc(3,1,
       "본인(현장대리인)은 당사 근로자에 대한 산업안전보건법상의 안전보건교육을 철저히 이행하였음을 확인합니다.",
       font=ft_sub, align=al)
    mg(3,20,3,23); bc(3,20,"현장대리인 :", font=ft_sub, align=ac)
    mg(3,24,3,28); bc(3,24, manager,       font=ft_sub, align=ac)
    mg(3,29,3,30); bc(3,29,"(인)",          font=ft_sub, align=ac)

    # ════════════════════════════════════════════════════
    # 헤더 3단 (행4=대분류, 행5=중분류, 행6=소분류)
    # ════════════════════════════════════════════════════

    # ── 대분류 (행4)
    big = [
        (4,1, 6,1,  "NO",                                             fill_gray),
        (4,2, 5,2,  "성명",                                            fill_hdr),
        (4,3, 4,5,  "업무형태 구분 Check:",                             fill_hdr),
        (4,6, 6,6,  "채용일자\n(최초투입일)",                           fill_hdr),
        (4,7, 4,8,  "채용·시교육이수현황 Check:",                        fill_hdr),
        (4,9, 4,11, "근로자 특별안전·보건교육\n[일용(단기&간헐): 2시간, 그 외: 16시간]", fill_spec),
        (4,12,4,23, "정기안전보건교육 이수현황 기입란\n[각 월 기입란에 교육이수 시간 기입]", fill_white),
        (4,24,4,25, "관리감독자\n연간 16시간",                          fill_mgr),
        (4,26,4,29, "작업(근로)자\n매반기 12시간",                      fill_wkr),
        (4,30,6,30, "비 고",                                         fill_gray),
    ]
    for r1,c1,r2,c2,text,fill in big:
        if not (r1==r2 and c1==c2): mg(r1,c1,r2,c2)
        bc(r1,c1, val=text, font=ft_bold, fill=fill, align=ac)

    # ── 중분류 (행5)
    mid = [
        (6,2, 6,2,  "생년월일",                  fill_hdr),
        (5,3, 6,3,  "관리감독자",                fill_hdr),
        (5,4, 5,5,  "작업(근로)자",              fill_hdr),
        (6,4, 6,4,  "상용",                     fill_hdr),
        (6,5, 6,5,  "일용",                     fill_hdr),
        (5,7, 6,7,  "일용\n1시간",              fill_hdr),
        (5,8, 6,8,  "일용 외\n8시간",           fill_hdr),
        (5,9, 6,9,  "대상여부\nCheck",          fill_spec),
        (5,10, 6,10, "대상\n작업명",            fill_spec),
        (5,11, 6,11, "교육이수일\n[이수시간]",   fill_spec),
        (5,12, 5,14, "1분기 (시간)",            fill_q1),
        (5,15, 5,17, "2분기 (시간)",            fill_q2),
        (5,18, 5,20, "3분기 (시간)",            fill_q3),
        (5,21, 5,23, "4분기 (시간)",            fill_q4),
        (6,12, 6,12, "1월",   fill_q1),
        (6,13, 6,13, "2월",   fill_q1),
        (6,14, 6,14, "3월",   fill_q1),
        (6,15, 6,15, "4월",   fill_q2),
        (6,16, 6,16, "5월",   fill_q2),
        (6,17, 6,17, "6월",   fill_q2),
        (6,18, 6,18, "7월",   fill_q3),
        (6,19, 6,19, "8월",   fill_q3),
        (6,20, 6,20, "9월",   fill_q3),
        (6,21, 6,21, "10월",  fill_q4),
        (6,22, 6,22, "11월",  fill_q4),
        (6,23, 6,23, "12월",  fill_q4),
        (5,24, 6,24, "매반",   fill_mgr),
        (5,25, 6,25, "3월",    fill_mgr),
        (5,26, 6,26, "매반",   fill_wkr),
        (5,27, 6,27, "6월",    fill_wkr),
        (5,28, 6,28, "매반",   fill_wkr),
        (5,29, 6,29, "12월",   fill_wkr),
    ]
    for r1,c1,r2,c2,text,fill in mid:
        if not (r1==r2 and c1==c2): mg(r1,c1,r2,c2)
        bc(r1,c1, val=text, font=ft_bold, fill=fill, align=ac)


    apply_bdr(4,1,6,LAST_COL)

    # ════════════════════════════════════════════════════
    # 데이터 행 (1인당 2행: 상=성명, 하=생년월일)
    # ════════════════════════════════════════════════════
    MAX_PERSONS = max(10, len(records))
    DATA_START  = 7

    month_col = {
        "q1_1":12,"q1_2":13,"q1_3":14,
        "q2_4":15,"q2_5":16,"q2_6":17,
        "q3_7":18,"q3_8":19,"q3_9":20,
        "q4_10":21,"q4_11":22,"q4_12":23,
    }
    month_fill_map = {
        12:fill_q1,13:fill_q1,14:fill_q1,
        15:fill_q2,16:fill_q2,17:fill_q2,
        18:fill_q3,19:fill_q3,20:fill_q3,
        21:fill_q4,22:fill_q4,23:fill_q4,
    }

    for i in range(MAX_PERSONS):
        ra = DATA_START + i*2
        rb = DATA_START + i*2 + 1
        ws.row_dimensions[ra].height = 35
        ws.row_dimensions[rb].height = 30
        rec = records[i] if i < len(records) else {}

        t = rec.get("type", "")
        chk_wkr   = "✔" if "근로" in t or t in ("근로자",) else ""
        chk_mgr_s = "✔" if "관리" in t and "상용" in t else (
                    "✔" if t in ("관리감독자", "관리") else "")
        chk_mgr_i = "✔" if "관리" in t and "임용" in t else ""

        # NO (2행 병합)
        mg(ra,1,rb,1); bc(ra,1, i+1, font=ft_normal, align=ac)
        # 성명(상) / 생년월일(하)
        bc(ra,2, rec.get("name",""),  font=ft_normal, align=ac)
        bc(rb,2, rec.get("birth",""), font=ft_small,  align=ac)
        # 업무형태 체크 (2행 병합)
        for c_, v_, f_ in [(3,chk_wkr,fill_hdr),(4,chk_mgr_s,fill_hdr),(5,chk_mgr_i,fill_hdr)]:
            mg(ra,c_,rb,c_); bc(ra,c_, v_, font=ft_bold, fill=f_, align=ac)
        # 채용일자 (2행 병합)
        mg(ra,6,rb,6); bc(ra,6, rec.get("hired",""), font=ft_small, align=ac)
        # 채용시교육 (2행 병합)
        mg(ra,7,rb,7); bc(ra,7, rec.get("채용시교육",""),    font=ft_small, align=ac)
        mg(ra,8,rb,8); bc(ra,8, rec.get("작업내용변경시",""), font=ft_small, align=ac)
        # 특별교육
        sp = rec.get("특별교육", "")
        sp_대상  = "✔" if sp in ("이수", "대상") else ""
        sp_비대상 = "✔" if sp == "비대상" else ""
        mg(ra,9,rb,9);   bc(ra,9,  sp_대상,  font=ft_bold, fill=fill_spec, align=ac)
        mg(ra,10,rb,10); bc(ra,10, sp_비대상, font=ft_bold, fill=fill_spec, align=ac)
        mg(ra,11,rb,11); bc(ra,11, rec.get("특별교육_이수일",""),
                           font=ft_small, fill=fill_spec, align=ac)
        # 월별 정기교육 (2행 병합)
        for mk, col in month_col.items():
            f = month_fill_map[col]
            mg(ra,col,rb,col)
            bc(ra,col, rec.get(mk,""), font=ft_small, fill=f, align=ac)
        # 관리감독자/작업자 연간 누계 (2행 병합, 빈칸)
        for col, f in [(24,fill_mgr),(25,fill_mgr),(26,fill_wkr),(27,fill_wkr),(28,fill_wkr),(29,fill_wkr)]:
            mg(ra,col,rb,col); bc(ra,col, "", font=ft_small, fill=f, align=ac)
        # 비고 (2행 병합)
        mg(ra,30,rb,30); bc(ra,30, "", font=ft_normal, align=al)

        apply_bdr(ra,1,rb,LAST_COL)

    # ── 하단 주석
    note_start = DATA_START + MAX_PERSONS*2 + 1
    notes = [
        "※ 상기 안전보건교육 이수 관리대장 작성 기준은 산업안전보건법상의 안전보건교육 이수를 의미함.",
        "※ 상기 교육이수 현황의 경우 산업안전보건법상의 교육실시 결과보고서(기록지)의 증빙이 가능한 교육이수자에 한하여 교육이수현황 기입할 것.",
        "※ 특별안전교육 시 2개월 이내 종료되는 일회성작업(단기간 작업) & 연간 총 작업일수 60일 미만작업(간헐적 작업)의 경우 2시간 교육으로 인정가능.",
    ]
    for j, note in enumerate(notes):
        nr = note_start + j
        ws.row_dimensions[nr].height = 35
        mg(nr,1,nr,LAST_COL)
        bc(nr,1, note, font=ft_small, align=al)

    last_note = note_start + len(notes) - 1
    ws.print_area = "A1:AD{}".format(last_note)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
