import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

output_path = r"C:\Users\-\OneDrive\바탕 화면\4.4.1_위험성평가표(콘테이너하차)_최종.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "콘테이너하차 위험성평가표"

# Styles
bold_font = Font(bold=True)
title_font = Font(bold=True, size=18)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

def set_border(ws, min_col, min_row, max_col, max_row):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border

for i in range(1, 23):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 4

# --- Header Section ---
ws.merge_cells('A1:U1')
ws['A1'] = "합동 위험성평가표 [ ■최초, □정기, □수시 ]"
ws['A1'].font = title_font
ws['A1'].alignment = center_align
ws.row_dimensions[1].height = 40

ws.merge_cells('A2:B3')
ws['A2'] = "작 업 명"
ws.merge_cells('C2:E3')
ws['C2'] = "가산~가평 천연가스 공급설비\n건설공사 비파괴검사 기술용역"
ws.merge_cells('F2:G2')
ws['F2'] = "평가기간"
ws.merge_cells('H2:M2')
ws['H2'] = "2026. 07. 09"
ws.merge_cells('N2:P2')
ws['N2'] = "재평가일"
ws.merge_cells('Q2:U2')

ws.merge_cells('A4:B5')
ws['A4'] = "회 사 명"
ws.merge_cells('C4:E5')
ws['C4'] = "서울검사(주)"

ws.merge_cells('A6:B7')
ws['A6'] = "작업공종"
ws.merge_cells('C6:E7')
ws['C6'] = "콘테이너 하차"

ws.merge_cells('A8:B9')
ws['A8'] = "작업기간"
ws.merge_cells('C8:E9')
ws['C8'] = "2026.06.20 ~ 2026.06.30"

ws.merge_cells('F3:F9')
ws['F3'] = "작업전"

ws.merge_cells('G3:G5')
ws['G3'] = "수급인"

ws.merge_cells('H3:I3')
ws['H3'] = "근 로 자"
ws.merge_cells('J3:K3')
ws['J3'] = "작 성 자"
ws.merge_cells('L3:M3')
ws['L3'] = "승 인 자"

ws.merge_cells('H4:I5')
ws['H4'] = "유상훈 (서명)"
ws.merge_cells('J4:K5')
ws['J4'] = "주진철 (서명)"
ws.merge_cells('L4:M5')
ws['L4'] = "강신태 (서명)"

ws.merge_cells('G6:G9')
ws['G6'] = "도급인"
ws.merge_cells('H6:I6')
ws['H6'] = "검토자(감독)"
ws.merge_cells('J6:K6')
ws['J6'] = "검토자(안전)"
ws.merge_cells('L6:M6')
ws['L6'] = "승 인 자"
ws.merge_cells('H7:I9')
ws['H7'] = "(서명)"
ws.merge_cells('J7:K9')
ws['J7'] = "(서명)"
ws.merge_cells('L7:M9')
ws['L7'] = "(서명)"

ws.merge_cells('N3:N9')
ws['N3'] = "재평가"

ws.merge_cells('O3:O5')
ws['O3'] = "수급인"

ws.merge_cells('P3:Q3')
ws['P3'] = "근 로 자"
ws.merge_cells('R3:S3')
ws['R3'] = "작 성 자"
ws.merge_cells('T3:U3')
ws['T3'] = "승 인 자"

ws.merge_cells('P4:Q5')
ws['P4'] = "(서명)"
ws.merge_cells('R4:S5')
ws['R4'] = "(서명)"
ws.merge_cells('T4:U5')
ws['T4'] = "(서명)"

ws.merge_cells('O6:O9')
ws['O6'] = "도급인"
ws.merge_cells('P6:Q6')
ws['P6'] = "검토자(감독)"
ws.merge_cells('R6:S6')
ws['R6'] = "검토자(안전)"
ws.merge_cells('T6:U6')
ws['T6'] = "승 인 자"
ws.merge_cells('P7:Q9')
ws['P7'] = "(서명)"
ws.merge_cells('R7:S9')
ws['R7'] = "(서명)"
ws.merge_cells('T7:U9')
ws['T7'] = "(서명)"

for r in range(2, 10):
    for c in range(1, 22):
        cell = ws.cell(row=r, column=c)
        cell.alignment = center_align
        cell.font = Font(size=10)

set_border(ws, 1, 2, 21, 9)

# --- Table Sub-Header ---
ws.merge_cells('A10:M10')
ws['A10'] = "작업 전 평가"
ws.merge_cells('N10:U10')
ws['N10'] = "재평가"

ws.merge_cells('A11:A12')
ws['A11'] = "세부작업"
ws.merge_cells('B11:E11')
ws['B11'] = "유해위험요인 파악"
ws.merge_cells('F11:G12')
ws['F11'] = "현재의 안전보건조치"
ws.merge_cells('H11:J11')
ws['H11'] = "위험성 결정"
ws.merge_cells('K11:O12')
ws['K11'] = "위험성 감소대책"
ws.merge_cells('P11:P12')
ws['P11'] = "개선\n예정일"
ws.merge_cells('Q11:Q12')
ws['Q11'] = "담당자\n(담당부서)"
ws.merge_cells('R11:R12')
ws['R11'] = "완료일"
ws.merge_cells('S11:U11')
ws['S11'] = "개선 후 위험성 (재평가)"

ws.merge_cells('B12:B12')
ws['B12'] = "위험\n분류"
ws.merge_cells('C12:E12')
ws['C12'] = "위험발생 상황 및 결과"
ws['H12'] = "가능성\n(빈도)"
ws['I12'] = "중대성\n(강도)"
ws['J12'] = "위험성\n(등급)"
ws['S12'] = "가능성\n(빈도)"
ws['T12'] = "중대성\n(강도)"
ws['U12'] = "위험성\n(등급)"

for r in range(10, 13):
    for c in range(1, 22):
        cell = ws.cell(row=r, column=c)
        cell.alignment = center_align
        cell.font = bold_font
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

set_border(ws, 1, 10, 21, 12)

# Column Widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 6
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 7 
ws.column_dimensions['I'].width = 7 
ws.column_dimensions['J'].width = 7 
ws.column_dimensions['K'].width = 7 
ws.column_dimensions['L'].width = 7 
ws.column_dimensions['M'].width = 7 
ws.column_dimensions['O'].width = 25 
ws.column_dimensions['P'].width = 7
ws.column_dimensions['Q'].width = 7
ws.column_dimensions['R'].width = 7

# --- Data Definition (Container Unloading) ---
data = [
    ["콘테이너\n하차", "기계\n(설비)적\n요인", "장비유도원 이동 중 이동식 크레인에 부딪힘, 끼임", "안전모, 안전화 착용/ 작업 전 점검", 2, 3, "V", "1. 크레인 이동 반경 내 통제선 설치 및 작업자 출입 통제\n2. 장비 전담 신호수(유도원) 배치 및 운전원과 무전 소통 확립"],
    ["콘테이너\n하차", "기계\n(설비)적\n요인", "이동식 크레인 반입 시 작업자와 충돌 위험", "안전모, 안전화 착용/ 작업 전 점검", 2, 3, "V", "1. 장비 진출입로 사전 확보 및 보행자 동선과 분리\n2. 반입 시 경광등 및 후진 알람 작동 확인"],
    ["콘테이너\n하차", "기계\n(설비)적\n요인", "슬링벨트를 불량하게 체결한 상태로 슬링벨트 이탈 낙하", "작업자 머리위로 위치하거나 지나지 않는다.", 1, 3, "VI", "1. 인양 전 슬링벨트 2줄 걸이 이상 견고한 결속 상태 교차 점검\n2. 인양 중인 중량물 하부 전면 통행 금지"],
    ["콘테이너\n하차", "기계\n(설비)적\n요인", "근로자가 안전모 등 개인보호구 미착용에 의한 충돌", "안전모, 안전화 착용/ 작업 전 점검", 2, 2, "V", "1. 작업장 진입 전 관리감독자의 개인보호구 착용 상태 점검\n2. 안전모 턱끈 결속 및 안전화 착용 철저"],
    ["콘테이너\n하차", "기계\n(설비)적\n요인", "중장비 이용 중량물 취급작업 중 신호오류에 의한 충돌사고 위험", "안전모, 안전화 착용/ 작업 전 점검", 1, 2, "V", "1. 지정된 신호수 1인의 수신호 및 무전에 의해서만 작업 지휘\n2. 운전자의 시야가 확보되지 않는 사각지대 작업 원칙적 금지"],
    ["콘테이너\n하차", "전기적\n요인", "인양 작업장 주변 고압선로 접촉으로 감전", "전압에 따른 접근한계거리 설정", 1, 3, "VI", "1. 작업 반경 내 가공 전선로 유무 사전 확인 및 법적 이격 거리 준수\n2. 필요 시 전선 보호관 씌우기 또는 절연 조치 후 작업 실시"],
    ["콘테이너\n하차", "작업특성\n요인", "카고크레인 작업 후 운전자가 조정석에서 하차 도중 바닥으로 추락", "승하강 손잡이 설치", 2, 1, "VI", "1. 장비 승하강 시 반드시 3점 지지(두 손, 한 발 등) 원칙 준수\n2. 승하강 발판 및 손잡이의 미끄럼(오일, 수분 등) 방지 청소"],
    ["콘테이너\n하차", "작업환경\n요인", "신호체계 미정립 상황에서 중장비 이동 및 작업중 사고 발생", "작업자와 운전자의 안전체계적립", 2, 3, "V", "1. 작업 전 TBM(툴박스미팅)을 통해 운전원-신호수 간 표준 신호체계 합의\n2. 동시 다발적 작업 금지 및 지휘 계통 일원화 철저"]
]

start_row = 13
current_row = 13

groups = {}
for item in data:
    cat = item[0]
    groups[cat] = groups.get(cat, 0) + 1

for item in data:
    ws.cell(row=current_row, column=1, value=item[0]).alignment = center_align
    ws.cell(row=current_row, column=2, value=item[1]).alignment = center_align
    
    ws.merge_cells(f'C{current_row}:E{current_row}')
    ws.cell(row=current_row, column=3, value=item[2]).alignment = left_align
    ws.merge_cells(f'F{current_row}:G{current_row}')
    ws.cell(row=current_row, column=6, value=item[3]).alignment = left_align
    ws.cell(row=current_row, column=8, value=item[4]).alignment = center_align
    ws.cell(row=current_row, column=9, value=item[5]).alignment = center_align
    ws.cell(row=current_row, column=10, value=item[6]).alignment = center_align
    
    ws.merge_cells(f'K{current_row}:O{current_row}')
    ws.cell(row=current_row, column=11, value=item[7]).alignment = left_align
    
    ws.row_dimensions[current_row].height = 40
    current_row += 1

r = 13
for cat, count in groups.items():
    if count > 1:
        ws.merge_cells(f'A{r}:A{r+count-1}')
    r += count

def merge_sub_categories():
    r_idx = 13
    while r_idx < current_row:
        val = ws.cell(row=r_idx, column=2).value
        count = 1
        for j in range(r_idx + 1, current_row):
            if ws.cell(row=j, column=2).value == val and ws.cell(row=j, column=1).value == ws.cell(row=r_idx, column=1).value:
                count += 1
            else:
                break
        if count > 1:
            ws.merge_cells(f'B{r_idx}:B{r_idx+count-1}')
        r_idx += count

merge_sub_categories()

set_border(ws, 1, 13, 21, current_row - 1)

try:
    wb.save(output_path)
    print(f"Excel file created at: {output_path}")
except PermissionError:
    alt_path = output_path.replace(".xlsx", "_1.xlsx")
    wb.save(alt_path)
    print(f"Excel file created at: {alt_path}")
