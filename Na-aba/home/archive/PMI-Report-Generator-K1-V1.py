import sys
import subprocess
import os
import glob
import math
import traceback
import re
import warnings
import json 
from tkinter import filedialog, Tk, Label, Entry, Button, Frame, StringVar, ttk
from tkinter import messagebox

# ========================================================
# [0] 필수 라이브러리 체크 및 설치
# ========================================================
def install_package(package):
    try:
        __import__(package)
    except ImportError:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        except:
            print(f"⚠️ {package} 설치 실패. 수동 설치가 필요할 수 있습니다.")

install_package('openpyxl')
install_package('pandas')
install_package('Pillow') 

import pandas as pd
import openpyxl 
import datetime 
from openpyxl.cell.cell import MergedCell 
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.image import Image as XLImage 
from openpyxl.worksheet.datavalidation import DataValidation 
from openpyxl.styles import Alignment  # ★ 줄바꿈 기능을 위해 필수
from PIL import Image as PILImage, ImageChops 

from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# 경고 메시지 무시
warnings.simplefilter("ignore")

# ========================================================
# [설정] 전역 변수 초기화
# ========================================================

# [1] 갑지 (Cover) 변수
SEOUL_COVER_ANCHOR="E4"; SEOUL_COVER_W=200; SEOUL_COVER_H=18; SEOUL_COVER_X=+5; SEOUL_COVER_Y=+3
SITCO_COVER_ANCHOR="A5"; SITCO_COVER_W=100; SITCO_COVER_H=60; SITCO_COVER_X=35; SITCO_COVER_Y=-15
FOOTER_COVER_ANCHOR="Q51"; FOOTER_COVER_W=100; FOOTER_COVER_H=20; FOOTER_COVER_X=10; FOOTER_COVER_Y=0
FOOTER_PT_COVER_ANCHOR="A51"; FOOTER_PT_COVER_W=100; FOOTER_PT_COVER_H=25; FOOTER_PT_COVER_X=8; FOOTER_PT_COVER_Y=-5

# [2] 을지 (Data) 변수
SEOUL_DATA_ANCHOR="E4"; SEOUL_DATA_W=200; SEOUL_DATA_H=18; SEOUL_DATA_X=+5; SEOUL_DATA_Y=+3
SITCO_DATA_ANCHOR="A5"; SITCO_DATA_W=100; SITCO_DATA_H=60; SITCO_DATA_X=35; SITCO_DATA_Y=-15
FOOTER_DATA_ANCHOR="R40"; FOOTER_DATA_W=100; FOOTER_DATA_H=20; FOOTER_DATA_X=10; FOOTER_DATA_Y=0
FOOTER_PT_DATA_ANCHOR="A40"; FOOTER_PT_DATA_W=100; FOOTER_PT_DATA_H=20; FOOTER_PT_DATA_X=8; FOOTER_PT_DATA_Y=-5

# [3] 행 설정
START_ROW = 20       
DATA_END_ROW = 41    
PRINT_END_ROW = 43   

PAGE_NUMBER_CELL = "P2" 

# 경로 설정
if getattr(sys, 'frozen', False): folder_path = os.path.dirname(sys.executable)
elif '__file__' in globals(): folder_path = os.path.dirname(os.path.abspath(__file__))
else: folder_path = os.getcwd()
logo_folder_path = folder_path 
SETTINGS_FILE = os.path.join(folder_path, "logo_settings.json")

# ========================================================
# [설정 저장/불러오기]
# ========================================================
def load_settings():
    global SEOUL_COVER_X, SEOUL_COVER_Y, SEOUL_COVER_W, SEOUL_COVER_H, SEOUL_COVER_ANCHOR
    global SITCO_COVER_X, SITCO_COVER_Y, SITCO_COVER_W, SITCO_COVER_H, SITCO_COVER_ANCHOR
    global FOOTER_COVER_X, FOOTER_COVER_Y, FOOTER_COVER_W, FOOTER_COVER_H, FOOTER_COVER_ANCHOR
    global FOOTER_PT_COVER_X, FOOTER_PT_COVER_Y, FOOTER_PT_COVER_W, FOOTER_PT_COVER_H, FOOTER_PT_COVER_ANCHOR
    
    global SEOUL_DATA_X, SEOUL_DATA_Y, SEOUL_DATA_W, SEOUL_DATA_H, SEOUL_DATA_ANCHOR
    global SITCO_DATA_X, SITCO_DATA_Y, SITCO_DATA_W, SITCO_DATA_H, SITCO_DATA_ANCHOR
    global FOOTER_DATA_X, FOOTER_DATA_Y, FOOTER_DATA_W, FOOTER_DATA_H, FOOTER_DATA_ANCHOR
    global FOOTER_PT_DATA_X, FOOTER_PT_DATA_Y, FOOTER_PT_DATA_W, FOOTER_PT_DATA_H, FOOTER_PT_DATA_ANCHOR
    
    global START_ROW, DATA_END_ROW, PRINT_END_ROW

    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 갑지
            SEOUL_COVER_ANCHOR = data.get('SEOUL_COVER_ANCHOR', SEOUL_COVER_ANCHOR)
            SEOUL_COVER_X = data.get('SEOUL_COVER_X', SEOUL_COVER_X)
            SEOUL_COVER_Y = data.get('SEOUL_COVER_Y', SEOUL_COVER_Y)
            SEOUL_COVER_W = data.get('SEOUL_COVER_W', SEOUL_COVER_W)
            SEOUL_COVER_H = data.get('SEOUL_COVER_H', SEOUL_COVER_H)

            SITCO_COVER_ANCHOR = data.get('SITCO_COVER_ANCHOR', SITCO_COVER_ANCHOR)
            SITCO_COVER_X = data.get('SITCO_COVER_X', SITCO_COVER_X)
            SITCO_COVER_Y = data.get('SITCO_COVER_Y', SITCO_COVER_Y)
            SITCO_COVER_W = data.get('SITCO_COVER_W', SITCO_COVER_W)
            SITCO_COVER_H = data.get('SITCO_COVER_H', SITCO_COVER_H)

            FOOTER_COVER_ANCHOR = data.get('FOOTER_COVER_ANCHOR', FOOTER_COVER_ANCHOR)
            FOOTER_COVER_X = data.get('FOOTER_COVER_X', FOOTER_COVER_X)
            FOOTER_COVER_Y = data.get('FOOTER_COVER_Y', FOOTER_COVER_Y)
            FOOTER_COVER_W = data.get('FOOTER_COVER_W', FOOTER_COVER_W)
            FOOTER_COVER_H = data.get('FOOTER_COVER_H', FOOTER_COVER_H)

            FOOTER_PT_COVER_ANCHOR = data.get('FOOTER_PT_COVER_ANCHOR', FOOTER_PT_COVER_ANCHOR)
            FOOTER_PT_COVER_X = data.get('FOOTER_PT_COVER_X', FOOTER_PT_COVER_X)
            FOOTER_PT_COVER_Y = data.get('FOOTER_PT_COVER_Y', FOOTER_PT_COVER_Y)
            FOOTER_PT_COVER_W = data.get('FOOTER_PT_COVER_W', FOOTER_PT_COVER_W)
            FOOTER_PT_COVER_H = data.get('FOOTER_PT_COVER_H', FOOTER_PT_COVER_H)

            # 을지
            SEOUL_DATA_ANCHOR = data.get('SEOUL_DATA_ANCHOR', SEOUL_DATA_ANCHOR)
            SEOUL_DATA_X = data.get('SEOUL_DATA_X', SEOUL_DATA_X)
            SEOUL_DATA_Y = data.get('SEOUL_DATA_Y', SEOUL_DATA_Y)
            SEOUL_DATA_W = data.get('SEOUL_DATA_W', SEOUL_DATA_W)
            SEOUL_DATA_H = data.get('SEOUL_DATA_H', SEOUL_DATA_H)

            SITCO_DATA_ANCHOR = data.get('SITCO_DATA_ANCHOR', SITCO_DATA_ANCHOR)
            SITCO_DATA_X = data.get('SITCO_DATA_X', SITCO_DATA_X)
            SITCO_DATA_Y = data.get('SITCO_DATA_Y', SITCO_DATA_Y)
            SITCO_DATA_W = data.get('SITCO_DATA_W', SITCO_DATA_W)
            SITCO_DATA_H = data.get('SITCO_DATA_H', SITCO_DATA_H)

            FOOTER_DATA_ANCHOR = data.get('FOOTER_DATA_ANCHOR', FOOTER_DATA_ANCHOR)
            FOOTER_DATA_X = data.get('FOOTER_DATA_X', FOOTER_DATA_X)
            FOOTER_DATA_Y = data.get('FOOTER_DATA_Y', FOOTER_DATA_Y)
            FOOTER_DATA_W = data.get('FOOTER_DATA_W', FOOTER_DATA_W)
            FOOTER_DATA_H = data.get('FOOTER_DATA_H', FOOTER_DATA_H)

            FOOTER_PT_DATA_ANCHOR = data.get('FOOTER_PT_DATA_ANCHOR', FOOTER_PT_DATA_ANCHOR)
            FOOTER_PT_DATA_X = data.get('FOOTER_PT_DATA_X', FOOTER_PT_DATA_X)
            FOOTER_PT_DATA_Y = data.get('FOOTER_PT_DATA_Y', FOOTER_PT_DATA_Y)
            FOOTER_PT_DATA_W = data.get('FOOTER_PT_DATA_W', FOOTER_PT_DATA_W)
            FOOTER_PT_DATA_H = data.get('FOOTER_PT_DATA_H', FOOTER_PT_DATA_H)

            START_ROW = data.get('START_ROW', START_ROW)
            DATA_END_ROW = data.get('DATA_END_ROW', DATA_END_ROW)
            PRINT_END_ROW = data.get('PRINT_END_ROW', PRINT_END_ROW)
            
            print("✅ 저장된 설정을 불러왔습니다.")
        except Exception as e:
            print(f"⚠️ 설정 불러오기 실패 (기본값 사용): {e}")

def save_settings():
    data = {
        'SEOUL_COVER_ANCHOR': SEOUL_COVER_ANCHOR, 'SEOUL_COVER_X': SEOUL_COVER_X, 'SEOUL_COVER_Y': SEOUL_COVER_Y, 'SEOUL_COVER_W': SEOUL_COVER_W, 'SEOUL_COVER_H': SEOUL_COVER_H,
        'SITCO_COVER_ANCHOR': SITCO_COVER_ANCHOR, 'SITCO_COVER_X': SITCO_COVER_X, 'SITCO_COVER_Y': SITCO_COVER_Y, 'SITCO_COVER_W': SITCO_COVER_W, 'SITCO_COVER_H': SITCO_COVER_H,
        'FOOTER_COVER_ANCHOR': FOOTER_COVER_ANCHOR, 'FOOTER_COVER_X': FOOTER_COVER_X, 'FOOTER_COVER_Y': FOOTER_COVER_Y, 'FOOTER_COVER_W': FOOTER_COVER_W, 'FOOTER_COVER_H': FOOTER_COVER_H,
        'FOOTER_PT_COVER_ANCHOR': FOOTER_PT_COVER_ANCHOR, 'FOOTER_PT_COVER_X': FOOTER_PT_COVER_X, 'FOOTER_PT_COVER_Y': FOOTER_PT_COVER_Y, 'FOOTER_PT_COVER_W': FOOTER_PT_COVER_W, 'FOOTER_PT_COVER_H': FOOTER_PT_COVER_H,
        
        'SEOUL_DATA_ANCHOR': SEOUL_DATA_ANCHOR, 'SEOUL_DATA_X': SEOUL_DATA_X, 'SEOUL_DATA_Y': SEOUL_DATA_Y, 'SEOUL_DATA_W': SEOUL_DATA_W, 'SEOUL_DATA_H': SEOUL_DATA_H,
        'SITCO_DATA_ANCHOR': SITCO_DATA_ANCHOR, 'SITCO_DATA_X': SITCO_DATA_X, 'SITCO_DATA_Y': SITCO_DATA_Y, 'SITCO_DATA_W': SITCO_DATA_W, 'SITCO_DATA_H': SITCO_DATA_H,
        'FOOTER_DATA_ANCHOR': FOOTER_DATA_ANCHOR, 'FOOTER_DATA_X': FOOTER_DATA_X, 'FOOTER_DATA_Y': FOOTER_DATA_Y, 'FOOTER_DATA_W': FOOTER_DATA_W, 'FOOTER_DATA_H': FOOTER_DATA_H,
        'FOOTER_PT_DATA_ANCHOR': FOOTER_PT_DATA_ANCHOR, 'FOOTER_PT_DATA_X': FOOTER_PT_DATA_X, 'FOOTER_PT_DATA_Y': FOOTER_PT_DATA_Y, 'FOOTER_PT_DATA_W': FOOTER_PT_DATA_W, 'FOOTER_PT_DATA_H': FOOTER_PT_DATA_H,
        
        'START_ROW': START_ROW, 
        'DATA_END_ROW': DATA_END_ROW, 
        'PRINT_END_ROW': PRINT_END_ROW
    }
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4)
        print("✅ 설정이 저장되었습니다.")
    except Exception as e:
        print(f"⚠️ 설정 저장 실패: {e}")

# ========================================================
# [UI 함수] 설정 팝업
# ========================================================
def select_folder_ui(title):
    root = Tk(); root.withdraw(); root.attributes('-topmost', True) 
    dir_path = filedialog.askdirectory(title=title, initialdir=folder_path)
    root.destroy(); return dir_path

def select_file_ui(title, filetypes):
    root = Tk(); root.withdraw(); root.attributes('-topmost', True) 
    file_path = filedialog.askopenfilename(title=title, initialdir=folder_path, filetypes=filetypes)
    root.destroy(); return file_path

def open_position_settings_ui():
    global SEOUL_COVER_ANCHOR, SEOUL_COVER_X, SEOUL_COVER_Y, SEOUL_COVER_W, SEOUL_COVER_H
    global SITCO_COVER_ANCHOR, SITCO_COVER_X, SITCO_COVER_Y, SITCO_COVER_W, SITCO_COVER_H
    global FOOTER_COVER_ANCHOR, FOOTER_COVER_X, FOOTER_COVER_Y, FOOTER_COVER_W, FOOTER_COVER_H
    global FOOTER_PT_COVER_ANCHOR, FOOTER_PT_COVER_X, FOOTER_PT_COVER_Y, FOOTER_PT_COVER_W, FOOTER_PT_COVER_H
    
    global SEOUL_DATA_ANCHOR, SEOUL_DATA_X, SEOUL_DATA_Y, SEOUL_DATA_W, SEOUL_DATA_H
    global SITCO_DATA_ANCHOR, SITCO_DATA_X, SITCO_DATA_Y, SITCO_DATA_W, SITCO_DATA_H
    global FOOTER_DATA_ANCHOR, FOOTER_DATA_X, FOOTER_DATA_Y, FOOTER_DATA_W, FOOTER_DATA_H
    global FOOTER_PT_DATA_ANCHOR, FOOTER_PT_DATA_X, FOOTER_PT_DATA_Y, FOOTER_PT_DATA_W, FOOTER_PT_DATA_H
    
    global START_ROW, DATA_END_ROW, PRINT_END_ROW

    win = Tk(); win.title("세부 설정 (기준 셀/위치/크기)"); win.geometry("600x750"); win.attributes('-topmost', True)
    
    tabControl = ttk.Notebook(win)
    tab1 = ttk.Frame(tabControl); tab2 = ttk.Frame(tabControl); tab3 = ttk.Frame(tabControl)
    tabControl.add(tab1, text='[갑지] Cover'); tabControl.add(tab2, text='[을지] Data'); tabControl.add(tab3, text='[행] Rows')
    tabControl.pack(expand=1, fill="both")
    
    entries = {}
    
    def create_section(parent, row_start, title, defaults, key_prefix):
        Label(parent, text=title, font=("bold", 11)).grid(row=row_start, column=0, columnspan=6, pady=(15, 5), sticky="w")
        Label(parent, text="기준 셀:", fg="blue").grid(row=row_start+1, column=0, sticky='e')
        var_a = StringVar(value=str(defaults[0])); Entry(parent, textvariable=var_a, width=8, bg="#f0f8ff").grid(row=row_start+1, column=1, padx=2); entries[f"{key_prefix}_anchor"] = var_a
        Label(parent, text="(예: E4)").grid(row=row_start+1, column=2, sticky='w')

        Label(parent, text="가로(X):").grid(row=row_start+2, column=0, sticky='e')
        var_x = StringVar(value=str(defaults[1])); Entry(parent, textvariable=var_x, width=8).grid(row=row_start+2, column=1, padx=2); entries[f"{key_prefix}_x"] = var_x
        Label(parent, text="세로(Y):").grid(row=row_start+2, column=2, sticky='e')
        var_y = StringVar(value=str(defaults[2])); Entry(parent, textvariable=var_y, width=8).grid(row=row_start+2, column=3, padx=2); entries[f"{key_prefix}_y"] = var_y
        
        Label(parent, text="너비(W):").grid(row=row_start+3, column=0, sticky='e')
        var_w = StringVar(value=str(defaults[3])); Entry(parent, textvariable=var_w, width=8).grid(row=row_start+3, column=1, padx=2); entries[f"{key_prefix}_w"] = var_w
        Label(parent, text="높이(H):").grid(row=row_start+3, column=2, sticky='e')
        var_h = StringVar(value=str(defaults[4])); Entry(parent, textvariable=var_h, width=8).grid(row=row_start+3, column=3, padx=2); entries[f"{key_prefix}_h"] = var_h

    # 갑지
    create_section(tab1, 0, "1. SITCO (왼쪽 상단)", (SITCO_COVER_ANCHOR, SITCO_COVER_X, SITCO_COVER_Y, SITCO_COVER_W, SITCO_COVER_H), 'cover_sitco')
    create_section(tab1, 4, "2. 서울검사 (오른쪽 상단)", (SEOUL_COVER_ANCHOR, SEOUL_COVER_X, SEOUL_COVER_Y, SEOUL_COVER_W, SEOUL_COVER_H), 'cover_seoul')
    create_section(tab1, 8, "3. 바닥글 오른쪽", (FOOTER_COVER_ANCHOR, FOOTER_COVER_X, FOOTER_COVER_Y, FOOTER_COVER_W, FOOTER_COVER_H), 'cover_footer')
    create_section(tab1, 12, "4. 바닥글 왼쪽", (FOOTER_PT_COVER_ANCHOR, FOOTER_PT_COVER_X, FOOTER_PT_COVER_Y, FOOTER_PT_COVER_W, FOOTER_PT_COVER_H), 'cover_pt')

    # 을지
    create_section(tab2, 0, "1. SITCO (왼쪽 상단)", (SITCO_DATA_ANCHOR, SITCO_DATA_X, SITCO_DATA_Y, SITCO_DATA_W, SITCO_DATA_H), 'data_sitco')
    create_section(tab2, 4, "2. 서울검사 (오른쪽 상단)", (SEOUL_DATA_ANCHOR, SEOUL_DATA_X, SEOUL_DATA_Y, SEOUL_DATA_W, SEOUL_DATA_H), 'data_seoul')
    create_section(tab2, 8, "3. 바닥글 오른쪽", (FOOTER_DATA_ANCHOR, FOOTER_DATA_X, FOOTER_DATA_Y, FOOTER_DATA_W, FOOTER_DATA_H), 'data_footer')
    create_section(tab2, 12, "4. 바닥글 왼쪽", (FOOTER_PT_DATA_ANCHOR, FOOTER_PT_DATA_X, FOOTER_PT_DATA_Y, FOOTER_PT_DATA_W, FOOTER_PT_DATA_H), 'data_pt')

    # 행 설정
    Label(tab3, text="데이터 행 범위 설정", font=("bold", 12)).pack(pady=20)
    frame_rows = Frame(tab3); frame_rows.pack()
    Label(frame_rows, text="시작 행 (Start):").grid(row=0, column=0, padx=5, pady=5)
    var_s = StringVar(value=str(START_ROW)); Entry(frame_rows, textvariable=var_s, width=10).grid(row=0, column=1, padx=5, pady=5); entries['start_row'] = var_s
    Label(frame_rows, text="데이터 종료 행 (Data End):").grid(row=1, column=0, padx=5, pady=5)
    var_d_e = StringVar(value=str(DATA_END_ROW)); Entry(frame_rows, textvariable=var_d_e, width=10).grid(row=1, column=1, padx=5, pady=5); entries['data_end_row'] = var_d_e
    Label(frame_rows, text="인쇄 영역 종료 행 (Print End):").grid(row=2, column=0, padx=5, pady=5)
    var_p_e = StringVar(value=str(PRINT_END_ROW)); Entry(frame_rows, textvariable=var_p_e, width=10).grid(row=2, column=1, padx=5, pady=5); entries['print_end_row'] = var_p_e
    Label(tab3, text="* 데이터 종료 행: 데이터가 채워지는 마지막 줄 (예: 41)\n* 인쇄 영역 종료 행: 인쇄가 끝나는 줄 (예: 43)", fg="gray", justify="left").pack(pady=10)

    def on_apply():
        global SEOUL_COVER_ANCHOR, SEOUL_COVER_X, SEOUL_COVER_Y, SEOUL_COVER_W, SEOUL_COVER_H
        global SITCO_COVER_ANCHOR, SITCO_COVER_X, SITCO_COVER_Y, SITCO_COVER_W, SITCO_COVER_H
        global FOOTER_COVER_ANCHOR, FOOTER_COVER_X, FOOTER_COVER_Y, FOOTER_COVER_W, FOOTER_COVER_H
        global FOOTER_PT_COVER_ANCHOR, FOOTER_PT_COVER_X, FOOTER_PT_COVER_Y, FOOTER_PT_COVER_W, FOOTER_PT_COVER_H
        
        global SEOUL_DATA_ANCHOR, SEOUL_DATA_X, SEOUL_DATA_Y, SEOUL_DATA_W, SEOUL_DATA_H
        global SITCO_DATA_ANCHOR, SITCO_DATA_X, SITCO_DATA_Y, SITCO_DATA_W, SITCO_DATA_H
        global FOOTER_DATA_ANCHOR, FOOTER_DATA_X, FOOTER_DATA_Y, FOOTER_DATA_W, FOOTER_DATA_H
        global FOOTER_PT_DATA_ANCHOR, FOOTER_PT_DATA_X, FOOTER_PT_DATA_Y, FOOTER_PT_DATA_W, FOOTER_PT_DATA_H
        
        global START_ROW, DATA_END_ROW, PRINT_END_ROW
        
        try:
            SITCO_COVER_ANCHOR=entries['cover_sitco_anchor'].get(); SITCO_COVER_X=float(entries['cover_sitco_x'].get()); SITCO_COVER_Y=float(entries['cover_sitco_y'].get()); SITCO_COVER_W=float(entries['cover_sitco_w'].get()); SITCO_COVER_H=float(entries['cover_sitco_h'].get())
            SEOUL_COVER_ANCHOR=entries['cover_seoul_anchor'].get(); SEOUL_COVER_X=float(entries['cover_seoul_x'].get()); SEOUL_COVER_Y=float(entries['cover_seoul_y'].get()); SEOUL_COVER_W=float(entries['cover_seoul_w'].get()); SEOUL_COVER_H=float(entries['cover_seoul_h'].get())
            FOOTER_COVER_ANCHOR=entries['cover_footer_anchor'].get(); FOOTER_COVER_X=float(entries['cover_footer_x'].get()); FOOTER_COVER_Y=float(entries['cover_footer_y'].get()); FOOTER_COVER_W=float(entries['cover_footer_w'].get()); FOOTER_COVER_H=float(entries['cover_footer_h'].get())
            FOOTER_PT_COVER_ANCHOR=entries['cover_pt_anchor'].get(); FOOTER_PT_COVER_X=float(entries['cover_pt_x'].get()); FOOTER_PT_COVER_Y=float(entries['cover_pt_y'].get()); FOOTER_PT_COVER_W=float(entries['cover_pt_w'].get()); FOOTER_PT_COVER_H=float(entries['cover_pt_h'].get())

            SITCO_DATA_ANCHOR=entries['data_sitco_anchor'].get(); SITCO_DATA_X=float(entries['data_sitco_x'].get()); SITCO_DATA_Y=float(entries['data_sitco_y'].get()); SITCO_DATA_W=float(entries['data_sitco_w'].get()); SITCO_DATA_H=float(entries['data_sitco_h'].get())
            SEOUL_DATA_ANCHOR=entries['data_seoul_anchor'].get(); SEOUL_DATA_X=float(entries['data_seoul_x'].get()); SEOUL_DATA_Y=float(entries['data_seoul_y'].get()); SEOUL_DATA_W=float(entries['data_seoul_w'].get()); SEOUL_DATA_H=float(entries['data_seoul_h'].get())
            FOOTER_DATA_ANCHOR=entries['data_footer_anchor'].get(); FOOTER_DATA_X=float(entries['data_footer_x'].get()); FOOTER_DATA_Y=float(entries['data_footer_y'].get()); FOOTER_DATA_W=float(entries['data_footer_w'].get()); FOOTER_DATA_H=float(entries['data_footer_h'].get())
            FOOTER_PT_DATA_ANCHOR=entries['data_pt_anchor'].get(); FOOTER_PT_DATA_X=float(entries['data_pt_x'].get()); FOOTER_PT_DATA_Y=float(entries['data_pt_y'].get()); FOOTER_PT_DATA_W=float(entries['data_pt_w'].get()); FOOTER_PT_DATA_H=float(entries['data_pt_h'].get())

            START_ROW = int(entries['start_row'].get())
            DATA_END_ROW = int(entries['data_end_row'].get())
            PRINT_END_ROW = int(entries['print_end_row'].get())
            
            save_settings()
            win.destroy()
        except ValueError: messagebox.showerror("오류", "숫자는 정확히 입력해주세요.")

    Button(win, text="설정 적용 및 시작 (Start)", command=on_apply, bg="#d0e8f2", font=("bold", 12), height=2, width=30).pack(pady=20)
    win.mainloop()

# ========================================================
# [함수] 로고 찾기 & 배치
# ========================================================
def find_image_smart(keyword, exclude_keyword=None):
    candidates = glob.glob(os.path.join(logo_folder_path, "*.*"))
    valid_extensions = ['.PNG', '.JPG', '.JPEG', '.BMP', '.GIF']
    for path in candidates:
        fname = os.path.basename(path).upper(); ext = os.path.splitext(path)[1].upper()
        if ext not in valid_extensions: continue 
        if keyword.upper() in fname:
            if exclude_keyword and exclude_keyword.upper() in fname: continue
            return path 
    return None

def place_image_freely(ws, img_path, anchor_cell_str, w, h, x_offset, y_offset):
    try:
        if not img_path or not os.path.exists(img_path): return
        
        original = PILImage.open(img_path).convert("RGBA")
        resized = original.resize((int(w), int(h)), PILImage.Resampling.LANCZOS)
        temp_name = f"temp_{os.path.basename(img_path)}"
        temp_full_path = os.path.join(logo_folder_path, temp_name)
        try: resized.save(temp_full_path)
        except: 
            temp_full_path = os.path.join(folder_path, temp_name)
            resized.save(temp_full_path)
        
        img = XLImage(temp_full_path); img.width = w; img.height = h
        col_str, row_num = coordinate_from_string(anchor_cell_str)
        col_idx = column_index_from_string(col_str) - 1; row_idx = row_num - 1 
        
        emu_x = int(x_offset * 9525); emu_y = int(y_offset * 9525)
        emu_w = int(w * 9525); emu_h = int(h * 9525)
        
        marker = AnchorMarker(col=col_idx, colOff=emu_x, row=row_idx, rowOff=emu_y)
        size = XDRPositiveSize2D(cx=emu_w, cy=emu_h)
        img.anchor = OneCellAnchor(_from=marker, ext=size)
        ws.add_image(img)
    except Exception as e: print(f"⚠️ 이미지 배치 에러: {e}")

def add_logos_to_sheet(ws, is_cover=False):
    if not logo_folder_path: return 
    try: ws._images = [] 
    except: pass
    
    if is_cover:
        s_sitco = (SITCO_COVER_ANCHOR, SITCO_COVER_W, SITCO_COVER_H, SITCO_COVER_X, SITCO_COVER_Y)
        s_seoul = (SEOUL_COVER_ANCHOR, SEOUL_COVER_W, SEOUL_COVER_H, SEOUL_COVER_X, SEOUL_COVER_Y)
        s_footer = (FOOTER_COVER_ANCHOR, FOOTER_COVER_W, FOOTER_COVER_H, FOOTER_COVER_X, FOOTER_COVER_Y)
        s_pt = (FOOTER_PT_COVER_ANCHOR, FOOTER_PT_COVER_W, FOOTER_PT_COVER_H, FOOTER_PT_COVER_X, FOOTER_PT_COVER_Y)
    else:
        s_sitco = (SITCO_DATA_ANCHOR, SITCO_DATA_W, SITCO_DATA_H, SITCO_DATA_X, SITCO_DATA_Y)
        s_seoul = (SEOUL_DATA_ANCHOR, SEOUL_DATA_W, SEOUL_DATA_H, SEOUL_DATA_X, SEOUL_DATA_Y)
        s_footer = (FOOTER_DATA_ANCHOR, FOOTER_DATA_W, FOOTER_DATA_H, FOOTER_DATA_X, FOOTER_DATA_Y)
        s_pt = (FOOTER_PT_DATA_ANCHOR, FOOTER_PT_DATA_W, FOOTER_PT_DATA_H, FOOTER_PT_DATA_X, FOOTER_PT_DATA_Y)

    path = find_image_smart("SITCO"); 
    if path: place_image_freely(ws, path, *s_sitco)
    path = find_image_smart("서울검사"); 
    if path: place_image_freely(ws, path, *s_seoul)
    path = find_image_smart("바닥글", exclude_keyword="PMI-1"); 
    if not path: path = find_image_smart("PMI", exclude_keyword="PMI-1")
    if path: place_image_freely(ws, path, *s_footer)
    
    path_left = None
    if is_cover:
        path_left = find_image_smart("PMI갑")
    if not path_left: path_left = find_image_smart("PMI-1")
    if not path_left: path_left = find_image_smart("PT")
    if path_left: place_image_freely(ws, path_left, *s_pt)

def cleanup_temp_files():
    for p in [logo_folder_path, folder_path]:
        if not p: continue
        temp_files = glob.glob(os.path.join(p, "temp_*.png"))
        for f in temp_files:
            try: os.remove(f)
            except: pass

# ========================================================
# [함수] 데이터 처리 및 엑셀 유틸
# ========================================================
def to_float(val):
    if pd.isna(val): return 0.0
    s = str(val).upper().replace("%", "").strip()
    if "<" in s or "ND" in s or s == "": return 0.0
    try: return float(s)
    except ValueError: return 0.0

def guess_grade(mo_val, original_grade_text):
    if original_grade_text and str(original_grade_text).strip() != "": return str(original_grade_text).strip()
    try: return "SS316" if float(mo_val) >= 1.5 else "SS304"
    except: return ""

def find_col(df, keywords, exclude_keywords=None):
    if exclude_keywords is None: exclude_keywords = []
    for col in df.columns:
        c_str = str(col).upper().strip()
        if any(ex in c_str for ex in exclude_keywords): continue
        if any(k in c_str for k in keywords): return col
    return None

def force_copy_footer(source_ws, target_ws):
    try:
        if source_ws.oddFooter: target_ws.oddFooter.left.text = source_ws.oddFooter.left.text
    except: pass 

def force_print_settings(source_ws, target_ws):
    try:
        target_ws.print_area = f'A1:U{PRINT_END_ROW}'
        target_ws.row_breaks.append(Break(id=PRINT_END_ROW))
        if target_ws.page_setup:
            target_ws.page_setup.fitToPage = True; target_ws.page_setup.fitToHeight = False; target_ws.page_setup.fitToWidth = 1      
        if source_ws.page_margins and target_ws.page_margins:
            target_ws.page_margins.left = source_ws.page_margins.left; target_ws.page_margins.right = source_ws.page_margins.right
            target_ws.page_margins.top = source_ws.page_margins.top; target_ws.page_margins.bottom = source_ws.page_margins.bottom
            target_ws.page_margins.header = source_ws.page_margins.header; target_ws.page_margins.footer = source_ws.page_margins.footer
    except: pass

def prepare_next_sheet(wb, source_sheet_idx, page_num):
    source_sheet = wb.worksheets[source_sheet_idx]; new_sheet = wb.copy_worksheet(source_sheet) 
    base_title = source_sheet.title.split('_')[0]; new_sheet.title = f"{base_title[:20]}_{page_num:03d}"
    new_sheet.sheet_view.tabSelected = False
    
    force_copy_footer(source_sheet, new_sheet)
    force_print_settings(source_sheet, new_sheet)
    add_logos_to_sheet(new_sheet, is_cover=False)
    
    # 데이터 영역 초기화 (7열~16열)
    for r in range(START_ROW, DATA_END_ROW + 1):
        for c in range(7, 17): 
            cell = new_sheet.cell(row=r, column=c)
            if not isinstance(cell, MergedCell): cell.value = None
            else:
                for rng in new_sheet.merged_cells.ranges:
                    if cell.coordinate in rng: new_sheet.cell(row=rng.min_row, column=rng.min_col).value = None; break
    return new_sheet

# ========================================================
# [핵심] 엑셀 데이터 채우기 (좌우 분할 & 줄바꿈 모드)
# ========================================================
def fill_excel_pagination(wb, data_list):
    # [설정] 스타일 정의
    wrap_style = Alignment(wrap_text=True, horizontal='center', vertical='center')
    center_style = Alignment(horizontal='center', vertical='center')

    current_sheet_idx = 1 if len(wb.worksheets) >= 2 else 0
    current_page_num = 1
    ws = wb.worksheets[current_sheet_idx]
    
    # 시트명/페이지번호/로고 설정
    ws.title = f"{ws.title[:20]}_{current_page_num:03d}"
    try: ws[PAGE_NUMBER_CELL] = f"{current_page_num:03d}"
    except: pass 
    
    is_cover = (current_sheet_idx == 0)
    add_logos_to_sheet(ws, is_cover=is_cover)
    force_print_settings(ws, ws)
    
    # ========================================================
    # [컬럼 매핑] H:Ni, I:Cr, J:Mn, K:Mo / L:Ni, M:Cr, N:Mn, O:Mo
    # ========================================================
    col_no = 7      # G열
    col_grade = 16  # P열

    # 첫 번째 데이터 (왼쪽) - Ni, Cr, Mn, Mo 순서
    map_1 = {
        'Ni': 8,   # H
        'Cr': 9,   # I
        'Mn': 10,  # J
        'Mo': 11   # K
    }
    
    # 두 번째 데이터 (오른쪽) - Ni, Cr, Mn, Mo 순서
    map_2 = {
        'Ni': 12,  # L
        'Cr': 13,  # M
        'Mn': 14,  # N
        'Mo': 15   # O
    }

    current_row = START_ROW 
    
    for i in range(0, len(data_list), 2):
        if current_row > DATA_END_ROW:
            current_page_num += 1
            ws = prepare_next_sheet(wb, current_sheet_idx, current_page_num)
            try: ws[PAGE_NUMBER_CELL] = f"{current_page_num:03d}"
            except: pass
            current_sheet_idx += 1
            current_row = START_ROW 
        
        item1 = data_list[i]
        item2 = data_list[i+1] if (i + 1) < len(data_list) else None

        # (1) G열: No
        cell_no = ws.cell(row=current_row, column=col_no)
        if item2:
            cell_no.value = f"{item1['No']}\n{item2['No']}"
            cell_no.alignment = wrap_style
        else:
            cell_no.value = item1['No']
            cell_no.alignment = center_style

        # (2) H~K열: 첫 번째 데이터
        for key, col_idx in map_1.items():
            val = item1.get(key, 0.0)
            if val == 0 or val == 0.0: val = "" # [빈칸 처리]
            
            c = ws.cell(row=current_row, column=col_idx)
            c.value = val
            c.alignment = center_style

        # (3) L~O열: 두 번째 데이터
        if item2:
            for key, col_idx in map_2.items():
                val = item2.get(key, 0.0)
                if val == 0 or val == 0.0: val = "" # [빈칸 처리]
                
                c = ws.cell(row=current_row, column=col_idx)
                c.value = val
                c.alignment = center_style

        # (4) P열: Grade
        cell_grade = ws.cell(row=current_row, column=col_grade)
        g1 = str(item1.get('Grade', ''))
        
        if item2:
            g2 = str(item2.get('Grade', ''))
            if g1 == g2 and g1 != "":
                cell_grade.value = g1
                cell_grade.alignment = center_style
            else:
                cell_grade.value = f"{g1}\n{g2}"
                cell_grade.alignment = wrap_style
        else:
            cell_grade.value = g1
            cell_grade.alignment = center_style

        current_row += 1

    print(f"   ✅ 데이터 입력 완료 (좌우 분할 / 빈칸 처리 적용됨)")

# ========================================================
# [메인] 실행
# ========================================================
try:
    print("\n" + "="*30)
    print("[1] 로고 파일 확인 및 설정")
    logo_sitco = find_image_smart("SITCO"); logo_seoul = find_image_smart("서울검사")
    if not logo_sitco or not logo_seoul:
        print("👉 'SITCO'와 '서울검사' 로고가 들어있는 폴더를 선택해주세요.")
        selected_dir = select_folder_ui("로고 이미지 파일이 있는 폴더를 선택하세요")
        if selected_dir: logo_folder_path = selected_dir
    load_settings(); open_position_settings_ui()

    print("\n[2] 데이터 파일(RFI) 선택"); 
    target_file = select_file_ui("분석할 원본 데이터 선택 (RFI)", [("Excel files", "*.xls;*.xlsx;*.xlsm")])
    if not target_file: sys.exit()
    print(f"📂 선택됨: {os.path.basename(target_file)}")

    print("\n[3] 양식 파일(SIT-PMI) 선택"); 
    template_path = select_file_ui("SIT-PMI 양식 파일 선택", [("Excel files", "*.xlsx")])
    if not template_path: sys.exit()
    print(f"📂 선택됨: {os.path.basename(template_path)}")

    print("\n[4] 순번 선택 (선택 사항)"); 
    target_input = input("👉 특정 순번(NO)만 하려면 입력 (전체는 엔터): ").strip()
    target_no_list = []
    if target_input: target_no_list = [x.strip() for x in target_input.replace(',', ' ').split() if x.strip()]

    all_extracted_data = []
    xls = pd.ExcelFile(target_file)
    for sheet_name in xls.sheet_names:
        print(f"   [Processing] {sheet_name}...")
        try: temp_df = pd.read_excel(target_file, sheet_name=sheet_name, header=None, nrows=50)
        except: continue
        header_idx = None
        for i, row in temp_df.iterrows():
            row_str = str(row.values).upper()
            if ("CR" in row_str and "NI" in row_str) or ("CHROMIUM" in row_str): header_idx = i; break
        
        if header_idx is None:
            print(f"     -> ⚠️ 헤더 자동 검색 실패. 수동 입력 필요."); 
            try: print(temp_df.head(20).to_string()) 
            except: print(temp_df.head(20))
            user_input = input(f"👉 '{sheet_name}' 시트의 제목 줄 번호를 입력하세요 (Skip: Enter): ").strip()
            if user_input.isdigit(): header_idx = int(user_input)
            else: continue

        df = pd.read_excel(target_file, sheet_name=sheet_name, header=header_idx)
        col_cr = find_col(df, ["CR", "CHROMIUM"]); col_ni = find_col(df, ["NI", "NICKEL"])
        
        if not (col_cr and col_ni):
            print("\n⚠️ Cr/Ni 컬럼 인식 실패. 수동 지정 필요.")
            for idx, col_name in enumerate(df.columns): print(f"  [{idx}] {col_name}")
            try:
                idx_cr = input("👉 'Cr' 컬럼 번호 (Skip:Enter): ").strip()
                idx_ni = input("👉 'Ni' 컬럼 번호 (Skip:Enter): ").strip()
                idx_mo = input("👉 'Mo' 컬럼 번호 (선택): ").strip()
                if idx_cr.isdigit() and idx_ni.isdigit():
                    col_cr = df.columns[int(idx_cr)]; col_ni = df.columns[int(idx_ni)]
                    col_mo = df.columns[int(idx_mo)] if idx_mo.isdigit() else None
                else: continue
            except: continue

        col_mo = col_mo if 'col_mo' in locals() and col_mo else find_col(df, ["MO", "MOLYBDENUM"])
        col_mn = find_col(df, ["MN", "MANGANESE"])
        col_no = find_col(df, ["NO.", "NO", "SEQ", "NUM", "POS", "ITEM"]) 
        col_grade_orig = find_col(df, ["GRADE", "MATERIAL", "SPEC", "TYPE"])

        df['val_Cr'] = df[col_cr].apply(to_float); df['val_Ni'] = df[col_ni].apply(to_float)
        df['val_Mo'] = df[col_mo].apply(to_float) if col_mo else 0.0
        df['val_Mn'] = df[col_mn].apply(to_float) if col_mn else 0.0
        if col_no: df['val_No'] = df[col_no].astype(str).str.strip()
        else: df['val_No'] = range(1, len(df) + 1)
        df['val_Grade_Orig'] = df[col_grade_orig] if col_grade_orig else ""

        count_in_sheet = 0
        for _, row in df.iterrows():
            current_no = str(row['val_No'])
            if target_no_list and current_no not in target_no_list: continue 
            if row['val_Cr'] > 0 or (str(row['val_No']) != "" and str(row['val_No']) != "nan"):
                all_extracted_data.append({
                    'No': row['val_No'], 'Cr': row['val_Cr'], 'Ni': row['val_Ni'], 
                    'Mo': row['val_Mo'], 'Mn': row['val_Mn'], 'Grade': guess_grade(row['val_Mo'], row['val_Grade_Orig'])
                })
                count_in_sheet += 1
        print(f"     -> {count_in_sheet}개 데이터 확보")

    if not all_extracted_data: print("\n❌ 추출된 데이터가 없습니다."); sys.exit()

    now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = os.path.splitext(os.path.basename(template_path))[0]
    new_filename = f"{base_name}_완료_{now_str}.xlsx"
    new_save_path = os.path.join(logo_folder_path, new_filename)

    try:
        wb = openpyxl.load_workbook(template_path)
        if len(wb.worksheets) >= 1: add_logos_to_sheet(wb.worksheets[0], is_cover=True)
        fill_excel_pagination(wb, all_extracted_data)
        for sheet in wb.worksheets: sheet.sheet_view.tabSelected = False
        wb.active = 0; wb.worksheets[0].sheet_view.tabSelected = True
        wb.save(new_save_path)
        print(f"\n💾 완료! 저장된 파일: {new_filename}"); print(f"📂 폴더: {logo_folder_path}")
    except PermissionError: print(f"\n🚫 [오류] 엑셀 파일이 열려있습니다. 닫고 다시 시도하세요.")
    except Exception as e: print(f"\n🚫 오류: {e}"); traceback.print_exc()

except Exception as e: traceback.print_exc(); print("오류 발생")
finally: cleanup_temp_files(); input("\n종료하려면 엔터...")