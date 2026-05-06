import sys
import os
import glob
import math
import traceback
import re
import warnings
import json
import tempfile
import datetime
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, Border, Side
from PIL import Image as PILImage, ImageChops
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# Ignore library warnings
warnings.simplefilter("ignore")

# --- Constants & Paths ---
NAN_PATTERN = re.compile(r'^nan(\.0+)?$|^none$|^null$|^0\.0+|-0\.0+$', re.IGNORECASE)
DOT_ZERO_PATTERN = re.compile(r'\.0$')

# 공통 스타일 (테두리 등)
thin_side = Side(style='thin')
medium_side = Side(style='medium')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

if getattr(sys, 'frozen', False):
    SCRIPT_HOME = os.path.dirname(sys.executable)
else:
    SCRIPT_HOME = os.path.dirname(os.path.abspath(__file__))

SETTINGS_FILE = os.path.join(SCRIPT_HOME, "logo_settings_unified.json")

class NDTTab:
    def __init__(self, parent_notebook, tab_name, root):
        self.root = root
        self.tab_name = tab_name
        self.frame = tk.Frame(parent_notebook, background="#f9fafb")
        parent_notebook.add(self.frame, text=f" {tab_name} ")
        
        # --- State Variables ---
        self.logo_folder_path = tk.StringVar(value=SCRIPT_HOME)
        self.target_file_path = tk.StringVar()
        self.template_file_path = tk.StringVar()
        self.sequence_filter = tk.StringVar()
        
        # Option variables for extracting
        self.extraction_mode = tk.StringVar(value="전체") if tab_name == "PMI" else None
        
        self.auto_verify = tk.BooleanVar(value=True) if tab_name == "PMI" else None
        
        # [NEW] 원소 함량 필터링용 상태 변수
        self.element_filters = [] # list of dict: {'key': StringVar, 'min': StringVar, 'max': StringVar}
        if tab_name == "PMI":
            # 기본 필터 추가
            for k in ["Cr", "Ni", "Mo"]:
                self.element_filters.append({
                    'key': tk.StringVar(value=k),
                    'min': tk.StringVar(),
                    'max': tk.StringVar()
                })
        self.show_selected_only = tk.BooleanVar(value=False)
        self.extracted_data = [] # 현재 추출된 데이터 저장용
        
        # Default Settings - Different heights/anchors can be applied per tab later
        self.config = {
            # 갑지 (Cover)
            'SEOUL_COVER_ANCHOR': "E5", 'SEOUL_COVER_W': 200.0, 'SEOUL_COVER_H': 18.0, 'SEOUL_COVER_X': 30.0, 'SEOUL_COVER_Y': 15.0, 'SEOUL_COVER_PATH': "",
            'SITCO_COVER_ANCHOR': "A6", 'SITCO_COVER_W': 80.0, 'SITCO_COVER_H': 40.0, 'SITCO_COVER_X': 15.0, 'SITCO_COVER_Y': 10.0, 'SITCO_COVER_PATH': "",
            'FOOTER_COVER_ANCHOR': "Q51", 'FOOTER_COVER_W': 80.0, 'FOOTER_COVER_H': 20.0, 'FOOTER_COVER_X': -10.0, 'FOOTER_COVER_Y': 5.0, 'FOOTER_COVER_PATH': "",
            'FOOTER_PT_COVER_ANCHOR': "A51", 'FOOTER_PT_COVER_W': 100.0, 'FOOTER_PT_COVER_H': 25.0, 'FOOTER_PT_COVER_X': 3.0, 'FOOTER_PT_COVER_Y': 5.0, 'FOOTER_PT_COVER_PATH': "",
            'SKETCH_UP_COVER_ANCHOR': "A20", 'SKETCH_UP_COVER_W': 300.0, 'SKETCH_UP_COVER_H': 150.0, 'SKETCH_UP_COVER_X': 0.0, 'SKETCH_UP_COVER_Y': 0.0, 'SKETCH_UP_COVER_PATH': "",
            'SKETCH_DOWN_COVER_ANCHOR': "A30", 'SKETCH_DOWN_COVER_W': 300.0, 'SKETCH_DOWN_COVER_H': 150.0, 'SKETCH_DOWN_COVER_X': 0.0, 'SKETCH_DOWN_COVER_Y': 0.0, 'SKETCH_DOWN_COVER_PATH': "",
            # 을지 (Data)
            'SEOUL_DATA_ANCHOR': "F5", 'SEOUL_DATA_W': 200.0, 'SEOUL_DATA_H': 18.0, 'SEOUL_DATA_X': 35.0, 'SEOUL_DATA_Y': 15.0, 'SEOUL_DATA_PATH': "",
            'SITCO_DATA_ANCHOR': "A6", 'SITCO_DATA_W': 80.0, 'SITCO_DATA_H': 40.0, 'SITCO_DATA_X': 5.0, 'SITCO_DATA_Y': 10.0, 'SITCO_DATA_PATH': "",
            'FOOTER_DATA_ANCHOR': "Q37", 'FOOTER_DATA_W': 100.0, 'FOOTER_DATA_H': 15.0, 'FOOTER_DATA_X': 5.0, 'FOOTER_DATA_Y': 3.0, 'FOOTER_DATA_PATH': "",
            'FOOTER_PT_DATA_ANCHOR': "A37", 'FOOTER_PT_DATA_W': 100.0, 'FOOTER_PT_DATA_H': 30.0, 'FOOTER_PT_DATA_X': 3.0, 'FOOTER_PT_DATA_Y': -10.0, 'FOOTER_PT_DATA_PATH': "",
            # 행 설정
            'START_ROW': 11 if tab_name == "RT" else 19, 'DATA_END_ROW': 45, 'PRINT_END_ROW': 47
        }
        
        self.load_settings()
        
        # --- NDT Type Configuration ---
        self.columns_config = self._get_columns_for_mode(tab_name)
        self.column_keys = [c[0] for c in self.columns_config]
        self.display_columns = [c[1] for c in self.columns_config]
        self.column_widths = {c[1]: c[2] for c in self.columns_config}
        
        self.create_widgets()
        self.log(f"[INFO] {self.tab_name} 탭을 초기화했습니다.")
        
    def log(self, message):
        if hasattr(self, 'status_log'):
            self.status_log.config(state='normal')
            self.status_log.insert(tk.END, f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}\n")
            self.status_log.see(tk.END)
            self.status_log.config(state='disabled')
            self.root.update_idletasks()
        else:
            print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}")

    def load_settings(self):
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    saved_data = json.load(f)
                    self.config.update(saved_data)
                
                # [NEW] 탭별 전용 행 설정이 있으면 덮어씌우고, 없으면 생성자에서 설정한 탭별 기본값을 유지함
                # (단순 update 시 PMI의 19행이 RT를 덮어씌우는 것 방지)
                # RT는 항상 11행, PMI는 항상 19행이 기본값이 되도록 강제함
                row_defaults = {'START_ROW': 11 if self.tab_name == "RT" else 19, 'DATA_END_ROW': 45, 'PRINT_END_ROW': 47}
                for row_key in ['START_ROW', 'DATA_END_ROW', 'PRINT_END_ROW']:
                    tab_specific_key = f"{self.tab_name}_{row_key}"
                    if tab_specific_key in saved_data:
                        self.config[row_key] = saved_data[tab_specific_key]
                    else:
                        # 탭별 전용 설정이 파일에 없으면 하드코딩된 기본값 적용
                        self.config[row_key] = row_defaults[row_key]
                
                # [NEW] 저장된 필터 불러오기
                filter_key = f"{self.tab_name}_FILTERS"
                if filter_key in self.config:
                    saved_filters = self.config[filter_key]
                    if isinstance(saved_filters, list):
                        self.element_filters = []
                        for f_data in saved_filters:
                            self.element_filters.append({
                                'key': tk.StringVar(value=f_data.get('key', '')),
                                'min': tk.StringVar(value=f_data.get('min', '')),
                                'max': tk.StringVar(value=f_data.get('max', ''))
                            })
                print(f"SUCCESS: {self.tab_name} 설정을 불러왔습니다.")
            except Exception as e:
                print(f"WARNING: 설정 불러오기 실패 (기본값 사용): {e}")

    def save_settings(self):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                # Reload all data from file first so we don't accidentally overwrite other settings
                saved_data = json.load(f)
                self.config.update(saved_data)
                
            if hasattr(self, 'setting_vars'):
                for key, var in self.setting_vars.items():
                    val = var.get()
                    try:
                        if key.endswith(('_X', '_Y', '_W', '_H')): self.config[key] = float(val)
                        elif 'ROW' in key: self.config[key] = int(val)
                        else: self.config[key] = str(val)
                    except: pass

            self.config[f"{self.tab_name}_COLUMNS"] = self.columns_config
            
            # [NEW] 필터 설정 저장
            if self.tab_name == "PMI" and hasattr(self, 'element_filters'):
                filter_data = []
                for f_item in self.element_filters:
                    filter_data.append({
                        'key': f_item['key'].get(),
                        'min': f_item['min'].get(),
                        'max': f_item['max'].get()
                    })
                self.config[f"{self.tab_name}_FILTERS"] = filter_data

            # [NEW] 탭별 전용 행 설정 저장 (PMI 19행, RT 11행 등 독립 관리)
            for row_key in ['START_ROW', 'DATA_END_ROW', 'PRINT_END_ROW']:
                self.config[f"{self.tab_name}_{row_key}"] = self.config[row_key]

            with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=4)
            self.log("[SUCCESS] 설정이 파일에 저장되었습니다.")
        except Exception as e:
            self.log(f"[WARNING] 설정 저장 실패: {e}")

    def _get_columns_for_mode(self, mode):
        """Returns list of (data_key, display_name, width) for the Treeview depending on NDT mode."""
        custom_key = f"{mode}_COLUMNS"
        if custom_key in self.config:
            saved_cols = self.config[custom_key]
            if isinstance(saved_cols, list) and len(saved_cols) > 0:
                # [NEW] RT 탭의 경우, 이미지 기반 전면 개편(Film Ident No, Film Loc 등) 및 Welder No 추가를 위해 강제 업데이트
                # 지난 수정(20개 미만)보다 더 확실하게 'Film Ident' 라는 글자가 없거나 'Welder'가 없으면 초기화하도록 강화
                has_ident = any("Film Ident" in str(c) for c in saved_cols)
                has_welder = any("Welder" in str(c) for c in saved_cols)
                if mode == "RT" and (not has_ident or not has_welder): 
                    pass # Legacy settings found, proceed to use new defaults below
                else:
                    return [tuple(c) for c in saved_cols]

        base_cols = [
            ("selected", "V", 40), 
            ("No", "No", 50), 
            ("Date", "Date", 90), 
            ("Dwg", "Drawing No.", 180), 
            ("Joint", "Film Ident. No.", 100), 
            ("Loc", "Film Location", 100)
        ]
        if mode == "PMI":
            return base_cols + [("Ni", "Ni", 60), ("Cr", "Cr", 60), ("Mo", "Mo", 60), ("Grade", "Grade", 100)]
        elif mode == "RT":
            # RT 전용 전면 재구성 (이미지 기준)
            return [
                ("selected", "V", 40), 
                ("No", "No", 50), 
                ("Date", "Date", 90), 
                ("Dwg", "Drawing No.", 150), 
                ("Joint", "Film Ident. No.", 120), 
                ("Loc", "Film Location", 100),
                ("Accept", "Acc", 40), ("Reject", "Rej", 40), ("Grade", "Deg", 40),
                ("D1", "① Crack", 60), ("D2", "② IP", 60), ("D3", "③ LF", 60),
                ("D4", "④ Slag", 60), ("D5", "⑤ Por", 60), ("D6", "⑥ U/C", 60),
                ("D7", "⑦ RUC", 60), ("D8", "⑧ BT", 60), ("D9", "⑨ TI", 60),
                ("D10", "⑩ CP", 60), ("D11", "⑪ RC", 60), ("D12", "⑫ Mis", 60),
                ("D13", "⑬ EP", 60), ("D14", "⑭ SD", 60), ("D15", "⑮ Oth", 60),
                ("Welder", "Welder No", 100), ("Remarks", "Remarks", 120)
            ]
        elif mode == "PT":
            return base_cols + [("Result", "Result", 80)]
        elif mode == "MT":
            return base_cols + [("Result", "Result", 80)]
        elif mode == "PAUT":
            return base_cols + [("Defect", "Defect", 100), ("Result", "Result", 80)]
        return base_cols

    def create_widgets(self):
        # [FIX] Initialize style for ttk widgets
        style = ttk.Style()
        
        # --- Main Scrollable Container (FOR LOW RESOLUTION) ---
        self.canvas = tk.Canvas(self.frame, background="#f9fafb", highlightthickness=0, yscrollincrement=20)
        self.scrollbar = ttk.Scrollbar(self.frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, background="#f9fafb")

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas_frame_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        # Stretch internally: Keep frame width synced with canvas width
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfigure(self.canvas_frame_window, width=e.width))

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        # [NEW] 마우스 선택 영역 강제 시인성 (TEntry/Entry 공통)
        self.root.option_add('*selectBackground', '#3b82f6')
        self.root.option_add('*selectForeground', 'white')

        # Treeview Style for Grid Lines (Excel-like)
        style.configure("Treeview", rowheight=28, font=("Malgun Gothic", 10))
        # [CRITICAL] Ensure selected rows always use blue background even if tags are applied
        style.map("Treeview", 
                  background=[('selected', '#3b82f6')], 
                  foreground=[('selected', 'white')])
        style.configure("Treeview.Heading", font=("Malgun Gothic", 10, "bold"))

        # [REFINED] Focus-based Mouse Wheel Support
        def _on_mousewheel(event):
            # Only handle if this specific tab is currently visible
            if not self.frame.winfo_viewable():
                return
                
            try:
                focused = self.root.focus_get()
                curr = focused
                while curr:
                    try:
                        c_name = curr.winfo_class()
                        if c_name in ('Treeview', 'Text', 'Listbox'):
                            # For widgets, scroll faster by multiplying delta
                            factor = 2 if c_name == 'Treeview' else 1
                            curr.yview_scroll(int(-factor*(event.delta/120)), "units")
                            return "break"
                    except: pass
                    curr = curr.master
            except Exception: pass
            
            # Scroll the main canvas
            self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            return "break"
            
        # [CRITICAL] Use add="+" to allow multiple tabs to hook into MouseWheel
        self.root.bind_all("<MouseWheel>", _on_mousewheel, add="+")

        # [NEW] Focus-drop: Click on background to return focus to root (restores main scroll)
        def _on_root_click(event):
            try:
                # [CRITICAL] If clicking the Treeview or its children, DO NOT steal focus.
                # Stealing focus on Button-1 kills the drag-selection before it can start.
                if "treeview" in str(event.widget).lower():
                    return
                w_class = event.widget.winfo_class()
                if w_class in ('Frame', 'TFrame', 'Label', 'TLabel', 'Canvas', 'Labelframe', 'TLabelframe'):
                    self.root.focus_set()
            except: pass
        self.root.bind_all("<Button-1>", _on_root_click, add="+")

        # (Removed global focus-stealing click drop)
        main_container = tk.Frame(self.scrollable_frame, background="#f9fafb", padx=20, pady=20)
        main_container.pack(fill='both', expand=True)

        # Ensure scrollregion updates on all child changes
        def _on_child_configure(event):
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.scrollable_frame.bind("<Configure>", _on_child_configure, add="+")

        # [NEW] 스크롤 영역 끝단 보정용 더미 프레임
        tk.Frame(self.scrollable_frame, height=50, background="#f9fafb").pack(side='bottom', fill='x')

        tk.Label(main_container, text="PMI 성적서 자동 생성 및 검증 통합 도구 (V2)", font=("Malgun Gothic", 16, "bold"), background="#f9fafb", foreground="#111827").pack(pady=(0, 20), anchor='w')

        # 1. File Selection
        file_frame = ttk.LabelFrame(main_container, text=" 파일 및 폴더 선택 ", padding=15)
        file_frame.pack(fill='x', pady=(0, 20))

        def _add_file_row(parent, label, var, row, is_dir=False, types=None):
            ttk.Label(parent, text=label).grid(row=row, column=0, sticky='e', padx=5, pady=5)
            ttk.Entry(parent, textvariable=var, width=50, exportselection=False).grid(row=row, column=1, padx=5, pady=5)
            cmd = (lambda: self._browse_dir(var)) if is_dir else (lambda: self._browse_file(var, types))
            ttk.Button(parent, text="찾기", command=cmd).grid(row=row, column=2, padx=5, pady=5)

        _add_file_row(file_frame, "로고 폴더:", self.logo_folder_path, 0, is_dir=True)
        _add_file_row(file_frame, "RFI 데이터:", self.target_file_path, 1, types=[("Excel Source", "*.xls;*.xlsx;*.xlsm")])
        _add_file_row(file_frame, "성적서 양식:", self.template_file_path, 2, types=[("Excel Template", "*.xlsx;*.xlsm")])

        # 2. Configuration Tabs
        config_frame = ttk.LabelFrame(main_container, text=" 리포트 세부 설정 ", padding=5)
        config_frame.pack(fill='both', expand=False, pady=(0, 20))

        self.tab_notebook = ttk.Notebook(config_frame)
        self.tab_notebook.pack(fill='both', expand=True)

        self.setting_vars = {}
        tab_cover = ttk.Frame(self.tab_notebook, padding=10)
        tab_data = ttk.Frame(self.tab_notebook, padding=10)
        tab_rows = ttk.Frame(self.tab_notebook, padding=10)

        self.tab_notebook.add(tab_cover, text="갑지 (Cover)")
        self.tab_notebook.add(tab_data, text="을지 (Data)")
        self.tab_notebook.add(tab_rows, text="행 설정 (Rows)")

        # 2.4 Preview Tab (Added)
        self.tab_preview = ttk.Frame(self.tab_notebook, padding=5)
        self.tab_notebook.add(self.tab_preview, text="미리보기 (Preview)")
        self._create_preview_ui(self.tab_preview)

        self._create_setting_grid(tab_cover, "COVER")
        self._create_setting_grid(tab_data, "DATA")
        self._create_row_settings(tab_rows)

        # 3. Action Section (Integrated Options)
        action_outer = tk.Frame(main_container, background="#f9fafb")
        action_outer.pack(fill='x', pady=(0, 10))

        action_frame = tk.Frame(action_outer, background="#ffffff", highlightthickness=1, highlightbackground="#e5e7eb", padx=15, pady=15)
        action_frame.pack(fill='x')

        # Option: Sequence Filter
        filter_row = tk.Frame(action_frame, background="#ffffff")
        filter_row.pack(fill='x', pady=(0, 10))
        ttk.Label(filter_row, text="특정 순번(NO)만 추출 (예: 1, 3, 5-10):", background="#ffffff").pack(side='left', padx=(0, 10))
        ttk.Entry(filter_row, textvariable=self.sequence_filter, width=30, exportselection=False).pack(side='left')

        # Option: Auto Verify Toggle
        if self.auto_verify is not None:
            verify_row = tk.Frame(action_frame, background="#ffffff")
            verify_row.pack(fill='x', pady=(0, 10))
            tk.Checkbutton(verify_row, text="재질 자동 판정 (성분 비교 알고리즘 적용, 10% 허용오차)", variable=self.auto_verify, background="#ffffff", font=("Malgun Gothic", 10)).pack(side='left')
        
        # [NEW] 원소 함량 필터 UI (PMI 전용)
        if self.tab_name == "PMI":
            filter_ui_frame = tk.Frame(action_frame, background="#ffffff", pady=5)
            filter_ui_frame.pack(fill='x', pady=(0, 10))
            
            ttk.Label(filter_ui_frame, text="원소/컬럼 함량 필터 (Min~Max):", background="#ffffff", font=("Malgun Gothic", 10, "bold")).pack(side='left', padx=(0, 10))
            
            self.filter_container = tk.Frame(filter_ui_frame, background="#ffffff")
            self.filter_container.pack(side='left')
            
            def refresh_filter_ui():
                for widget in self.filter_container.winfo_children():
                    widget.destroy()
                
                for i, f_item in enumerate(self.element_filters):
                    unit_frame = tk.Frame(self.filter_container, background="#ffffff", padx=5)
                    unit_frame.pack(side='left')
                    
                    ttk.Entry(unit_frame, textvariable=f_item['key'], width=5, exportselection=False).pack(side='left')
                    ttk.Label(unit_frame, text=":", background="#ffffff").pack(side='left')
                    ttk.Entry(unit_frame, textvariable=f_item['min'], width=5, exportselection=False).pack(side='left')
                    ttk.Label(unit_frame, text="~", background="#ffffff").pack(side='left')
                    ttk.Entry(unit_frame, textvariable=f_item['max'], width=5, exportselection=False).pack(side='left')
                    
                    # 삭제 버튼
                    btn_del = tk.Button(unit_frame, text="×", command=lambda idx=i: remove_filter(idx), 
                                        relief='flat', background="#ffffff", foreground="red", font=("Arial", 10, "bold"))
                    btn_del.pack(side='left', padx=(2, 0))

            def add_filter():
                self.element_filters.append({'key': tk.StringVar(), 'min': tk.StringVar(), 'max': tk.StringVar()})
                refresh_filter_ui()

            def remove_filter(idx):
                if len(self.element_filters) > 0:
                    self.element_filters.pop(idx)
                    refresh_filter_ui()

            ttk.Button(filter_ui_frame, text="+ 필터 추가", command=add_filter, width=10).pack(side='left', padx=10)
            refresh_filter_ui()
        
        # Option: Extraction Mode (추출 방식)
        if self.extraction_mode is not None:
            mode_row = tk.Frame(action_frame, background="#ffffff")
            mode_row.pack(fill='x', pady=(0, 10))
            ttk.Label(mode_row, text="추출 방식 (Extraction Method):", background="#ffffff").pack(side='left', padx=(0, 10))
            mode_combo = ttk.Combobox(mode_row, textvariable=self.extraction_mode, state="readonly", width=25)
            mode_combo['values'] = ("전체", "SS304 만", "SS316 만", "DUPLEX 만", "SS310 만", "미분류(기타) 만")
            mode_combo.pack(side='left')

        # Buttons (Extract & Generate)
        btn_frame = tk.Frame(action_frame, background="#ffffff")
        btn_frame.pack(side='right')

        btn_extract = ttk.Button(btn_frame, text="데이터 추가 추출 (Add)", style="Action.TButton", command=self.extract_only)
        btn_extract.pack(side='left', padx=5)

        btn_start = ttk.Button(btn_frame, text="성적서 생성 시작 (Generate)", style="Action.TButton", command=self.run_process)
        btn_start.pack(side='left', padx=5)

        # 4. Progress bar
        self.progress = ttk.Progressbar(main_container, orient="horizontal", mode="determinate")
        self.progress.pack(fill='x', pady=(5, 15))

        # 5. Status Log Area
        log_frame = ttk.LabelFrame(main_container, text=" 작업 로그 (Status Log) ", padding=10)
        log_frame.pack(fill='both', expand=True)

        self.status_log = tk.Text(log_frame, height=8, font=("Consolas", 9), state='disabled', background="#000000", foreground="#10b981", padx=5, pady=5)
        vsb = ttk.Scrollbar(log_frame, orient="vertical", command=self.status_log.yview)
        self.status_log.configure(yscrollcommand=vsb.set)
        self.status_log.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        # [REMOVED] Conflicting focus/scroll overrides preserved standard OS behavior
        pass

    def _create_setting_grid(self, parent, context):
        items = [("SITCO 로고", f"SITCO_{context}"), ("서울검사 로고", f"SEOUL_{context}"), ("바닥글 우측", f"FOOTER_{context}"), ("바닥글 좌측", f"FOOTER_PT_{context}")]
        if self.tab_name == "RT" and context == "COVER":
            items.append(("Shooting Sketch (상)", f"SKETCH_UP_{context}"))
            items.append(("Shooting Sketch (하)", f"SKETCH_DOWN_{context}"))
        
        def pick_img(key):
            p = filedialog.askopenfilename(filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
            if p: self.setting_vars[key].set(p)

        for i, (label, key_prefix) in enumerate(items):
            # Label + Search Button
            lbl_frame = tk.Frame(parent, background="#ffffff")
            lbl_frame.grid(row=i, column=0, sticky='w', pady=(10, 2))
            ttk.Label(lbl_frame, text=label, font=("Malgun Gothic", 9, "bold")).pack(side='left')
            
            p_var = tk.StringVar(value=self.config.get(f"{key_prefix}_PATH", ""))
            self.setting_vars[f"{key_prefix}_PATH"] = p_var
            
            btn_search = tk.Button(lbl_frame, text="🔍", command=lambda k=f"{key_prefix}_PATH": pick_img(k), 
                                   relief='flat', background="#ffffff", font=("Arial", 9))
            btn_search.pack(side='left', padx=2)
            if p_var.get():
                btn_search.config(foreground="blue") # 파일이 지정됨을 표시
            
            ttk.Label(parent, text="셀:").grid(row=i, column=1, sticky='e')
            v_a = tk.StringVar(value=self.config.get(f"{key_prefix}_ANCHOR", ""))
            ttk.Entry(parent, textvariable=v_a, width=6, exportselection=False).grid(row=i, column=2, padx=2)
            self.setting_vars[f"{key_prefix}_ANCHOR"] = v_a
            for j, (coord, key_suffix) in enumerate([("X", "X"), ("Y", "Y"), ("W", "W"), ("H", "H")]):
                ttk.Label(parent, text=f"{coord}:").grid(row=i, column=3 + j*2, sticky='e')
                v = tk.StringVar(value=str(self.config.get(f"{key_prefix}_{key_suffix}", "0.0")))
                ttk.Entry(parent, textvariable=v, width=6, exportselection=False).grid(row=i, column=4 + j*2, padx=2)
                self.setting_vars[f"{key_prefix}_{key_suffix}"] = v

    def _create_row_settings(self, parent):
        rows = [("데이터 시작 행", "START_ROW", "실제 데이터가 입력되는 첫 행"), ("데이터 종료 행", "DATA_END_ROW", "한 페이지에서 데이터 입력이 끝나는 행"), ("인쇄 영역 종료 행", "PRINT_END_ROW", "페이지 구분선이 위치할 마지막 행")]
        for i, (label, key, tip) in enumerate(rows):
            ttk.Label(parent, text=label + ":").grid(row=i, column=0, sticky='e', padx=5, pady=8)
            v = tk.StringVar(value=str(self.config.get(key, 0)))
            ttk.Entry(parent, textvariable=v, width=10, exportselection=False).grid(row=i, column=1, sticky='w', padx=5)
            self.setting_vars[key] = v
            ttk.Label(parent, text=f"({tip})", foreground="gray", font=("Malgun Gothic", 9)).grid(row=i, column=2, sticky='w')

    def _create_preview_ui(self, parent):
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # Right: Sidebar buttons (Pack first to ensure visibility on the right)
        sidebar = tk.Frame(container, background="#f9fafb", padx=10)
        sidebar.pack(side='right', fill='y')

        # [NEW] Treeview와 수직/수평 스크롤바를 담을 내부 프레임
        tree_frame = tk.Frame(container, background="#f9fafb")
        tree_frame.pack(side='left', fill='both', expand=True)

        self.preview_tree = ttk.Treeview(tree_frame, columns=self.display_columns, show='headings', height=10, selectmode='extended')
        
        for col, width in self.column_widths.items():
            self.preview_tree.heading(col, text=col)
            self.preview_tree.column(col, width=width, anchor='center', stretch=False)
        
        # [MANUAL DRAG] Manually implement drag-selection since native is broken
        self.drag_start_item = None
        
        def _on_tree_press(event):
            self.preview_tree.focus_set()
            item_id = self.preview_tree.identify_row(event.y)
            column = self.preview_tree.identify_column(event.x)
            
            self.start_col = column
            self.end_col = self.start_col # [FIX] 클릭 시 종료 컬럼 초기화 (이전 드래그 잔상 제거)
            
            if item_id:
                # [NEW] 컬럼 인덱스 찾기
                try:
                    col_idx = int(column.replace("#", "")) - 1
                    if 0 <= col_idx < len(self.column_keys):
                        key = self.column_keys[col_idx]
                        view_idx = self.preview_tree.index(item_id)
                        
                        if hasattr(self, 'item_idx_map') and 0 <= view_idx < len(self.item_idx_map):
                            actual_idx = self.item_idx_map[view_idx]
                            
                            # (1) 선택 항목(체크박스) 토글
                            if key == "selected":
                                self.extracted_data[actual_idx]['selected'] = not self.extracted_data[actual_idx].get('selected', True)
                                self.populate_preview(self.extracted_data, switch_tab=False)
                                return "break"
                            
                            # (2) RT 결함 항목(D1~D15) 및 결과(Accept/Reject) 원클릭 토글
                            elif self.tab_name == "RT":
                                if (key.startswith("D") and key[1:].isdigit()) or (key in ["Accept", "Reject"]):
                                    old_v = self.extracted_data[actual_idx].get(key, "")
                                    new_v = "√" if old_v == "" else ""
                                    self.extracted_data[actual_idx][key] = new_v
                                    
                                    # Accept/Reject는 상호 배타적 선택 처리
                                    if key in ["Accept", "Reject"] and new_v == "√":
                                        other_key = "Reject" if key == "Accept" else "Accept"
                                        self.extracted_data[actual_idx][other_key] = ""
                                        
                                    self.populate_preview(self.extracted_data, switch_tab=False)
                                    return "break"
                except: pass

                self.drag_start_item = item_id
                if not (event.state & 0x0001 or event.state & 0x0004):
                    self.preview_tree.selection_set(item_id)

        def _on_tree_drag(event):
            if not self.drag_start_item: return
            curr_item = self.preview_tree.identify_row(event.y)
            if not curr_item: return
            self.end_col = self.preview_tree.identify_column(event.x)
            
            all_items = self.preview_tree.get_children('')
            try:
                idx_start = all_items.index(self.drag_start_item)
                idx_curr = all_items.index(curr_item)
                low = min(idx_start, idx_curr)
                high = max(idx_start, idx_curr)
                self.preview_tree.selection_set(all_items[low:high+1])
            except: pass

        def _on_tree_release(event):
            self.end_col = self.preview_tree.identify_column(event.x)
            self.drag_start_item = None
            self.on_tree_select(event)

        # [NEW] 수평/수직 스크롤바 추가 및 연결
        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.preview_tree.yview)
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        # [NEW] 그리드 배치로 스크롤바 정렬
        self.preview_tree.grid(row=0, column=0, sticky='nsew')
        scroll_y.grid(row=0, column=1, sticky='ns')
        scroll_x.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # (Rest of bindings...)
        self.preview_tree.bind("<Button-1>", _on_tree_press)
        self.preview_tree.bind("<B1-Motion>", _on_tree_drag)
        self.preview_tree.bind("<ButtonRelease-1>", _on_tree_release)
        self.preview_tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        self.preview_tree.bind("<Double-1>", self.on_tree_double_click)
        self.preview_tree.bind("<Button-3>", self.show_context_menu)
        
        # [NEW] 단축키 바인딩 (Copy/Paste/Delete)
        self.preview_tree.bind("<Control-c>", lambda e: self.copy_cell())
        self.preview_tree.bind("<Control-v>", lambda e: self.paste_cell())
        self.preview_tree.bind("<Delete>", lambda e: self.delete_item())
        
        # [NEW] 컬럼 너비 조절 자동 저장 바인딩
        def on_column_resize(event):
            # identify_region이 "separator"이면 너비를 조절한 것으로 간주
            region = self.preview_tree.identify_region(event.x, event.y)
            if region == "separator" or region == "heading":
                changed = False
                for i, (key, disp, old_w) in enumerate(self.columns_config):
                    curr_w = self.preview_tree.column(f"#{i+1}", "width")
                    if curr_w != old_w:
                        self.columns_config[i] = (key, disp, curr_w)
                        changed = True
                if changed:
                    self.save_settings()
                    self.log("📏 컬럼 너비가 자동 저장되었습니다.")

        self.preview_tree.bind("<ButtonRelease-1>", on_column_resize, add="+")

        ttk.Label(sidebar, text="데이터 관리", font=("Malgun Gothic", 9, "bold")).pack(pady=(0, 10))
        # [NEW] Date Filter Listbox
        ttk.Label(sidebar, text="날짜 선택 필터", font=("Malgun Gothic", 9, "bold")).pack(pady=(5, 2))
        list_frame = tk.Frame(sidebar)
        list_frame.pack(fill='x', pady=5)
        self.date_listbox = tk.Listbox(list_frame, selectmode='single', height=4, exportselection=False, 
                                       font=("Malgun Gothic", 9), selectbackground="#ffffff", 
                                       selectforeground="#000000", activestyle="none")
        self.date_listbox.pack(side='left', fill='both', expand=True)
        sb = ttk.Scrollbar(list_frame, orient='vertical', command=self.date_listbox.yview)
        sb.pack(side='right', fill='y')
        self.date_listbox.config(yscrollcommand=sb.set)
        
        def toggle_date(event):
            idx = self.date_listbox.nearest(event.y)
            if idx >= 0:
                val = self.date_listbox.get(idx)
                if val.startswith("[v]"):
                    self.date_listbox.delete(idx)
                    self.date_listbox.insert(idx, val.replace("[v] ", "[ ] "))
                elif val.startswith("[ ]"):
                    self.date_listbox.delete(idx)
                    self.date_listbox.insert(idx, val.replace("[ ] ", "[v] "))
                self.date_listbox.selection_clear(0, tk.END)
        
        self.date_listbox.bind("<ButtonRelease-1>", toggle_date)
        
        def apply_date_filter():
            selected_dates = [self.date_listbox.get(i).replace("[v] ", "").replace("[ ] ", "") for i in range(self.date_listbox.size()) if self.date_listbox.get(i).startswith("[v]")]
            for item in self.extracted_data:
                item['selected'] = (item.get('Date', '') in selected_dates)
            
            # Auto-check '선택 항목만 보기' when filter is applied
            self.show_selected_only.set(True)
            self.populate_preview(self.extracted_data, switch_tab=False)
            
        ttk.Button(sidebar, text="선택 날짜 적용", command=apply_date_filter).pack(fill='x', pady=(0, 10))
        
        
        # [NEW] View Selection Only Toggle
        ttk.Label(sidebar, text="항목 관리", font=("Malgun Gothic", 9, "bold")).pack(pady=(10, 2))
        ttk.Button(sidebar, text="행 추가 (Add Row)", width=12, command=self.add_manual_row).pack(pady=2)
        ttk.Button(sidebar, text="전체 선택", width=12, command=self.select_all).pack(pady=2)
        ttk.Button(sidebar, text="선택 해제", width=12, command=self.deselect_all).pack(pady=(2, 5))
        
        # [NEW] Column Settings Button
        ttk.Label(sidebar, text="설정 및 관리", font=("Malgun Gothic", 9, "bold")).pack(pady=(15, 2))
        ttk.Button(sidebar, text="컬럼(열) 설정", width=12, command=self.open_column_config_window).pack(pady=2)
        
        # [NEW] Column Settings Button
        ttk.Label(sidebar, text="▲ 순서 조절 ▼", font=("Malgun Gothic", 9, "bold")).pack(pady=(15, 2))
        ttk.Button(sidebar, text="위로 이동", width=12, command=lambda: self.move_item(-1)).pack(pady=2)
        ttk.Button(sidebar, text="아래로 이동", width=12, command=lambda: self.move_item(1)).pack(pady=2)

        tk.Frame(sidebar, height=1, background="#e5e7eb").pack(fill='x', pady=10) # Separator
        
        # [MOD] RT 탭에서는 사이드바의 '최종 실행' 버튼을 제거하여 UI를 간소화함 (사용자 요청)
        if self.tab_name != "RT":
            # [CRITICAL ACTION] Generate Report from Preview
            ttk.Label(sidebar, text="최종 실행", font=("Malgun Gothic", 9, "bold")).pack(pady=(5, 5))
            btn_gen = ttk.Button(sidebar, text="성적서 생성 시작", style="Action.TButton", width=15, command=self.run_process)
            btn_gen.pack(pady=5)
        
        tk.Frame(sidebar, height=10, background="#f9fafb").pack() # Spacer
        ttk.Button(sidebar, text="선택 삭제", width=12, command=self.delete_item).pack(pady=2)
        ttk.Button(sidebar, text="전체 초기화", width=12, command=self.clear_all).pack(pady=(2, 10))

        # [ContextMenu]
        self.ctx_menu = tk.Menu(self.root, tearoff=0)
        self.ctx_menu.add_command(label="셀 내용 복사 (Copy)", command=self.copy_cell)
        self.ctx_menu.add_command(label="셀 내용 붙여넣기 (Paste)", command=self.paste_cell)
        self.ctx_menu.add_command(label="선택 항목 복제 (Duplicate Row)", command=self.duplicate_selected_rows)
        self.ctx_menu.add_separator()
        self.ctx_menu.add_command(label="선택 항목 ISO 병합 (Merge ISO)", command=self.merge_selected_iso)
        self.ctx_menu.add_command(label="선택 항목 Joint 병합 (Merge Joint)", command=self.merge_selected_joint)

        # [NEW] 그룹 시인성을 위한 태그 색상 정의
        # [MOD] 엑셀과 같은 구분을 위해 색상 대비 강화
        self.preview_tree.tag_configure("evenrow", background="#f9fafb")
        self.preview_tree.tag_configure("oddrow", background="#ffffff")
        # [REFINED] Selection Visibility: Background colors are removed from tags to allow 
        # the native blue selection highlight to be clearly visible on Windows.
        for tag in ['group_even', 'group_odd', 'grouped_even', 'grouped_odd', 'item_even', 'item_odd', 'evenrow', 'oddrow']:
            self.preview_tree.tag_configure(tag, background="", foreground="black")
        
        # Use Bold for ISO group distinction instead of background colors
        self.preview_tree.tag_configure('group_even', font=("Malgun Gothic", 10, "bold"))
        self.preview_tree.tag_configure('group_odd', font=("Malgun Gothic", 10, "bold"))
        # [MOD] 연한 하늘색
        
        # [NEW] 모든 UI 생성이 끝난 후 Entry 위젯들을 찾아가서 우클릭 메뉴를 붙임
        self._create_entry_context_menu()

    def _create_entry_context_menu(self):
        self.entry_popup = tk.Menu(self.root, tearoff=0)
        self.entry_popup.add_command(label="복사 (Copy)", command=self._copy_entry)
        self.entry_popup.add_command(label="붙여넣기 (Paste)", command=self._paste_entry)
        self.entry_popup.add_separator()
        self.entry_popup.add_command(label="잘라내기 (Cut)", command=self._cut_entry)
        
        self.root.bind_class("Entry", "<Button-3>", self._show_entry_popup)
        self.root.bind_class("TEntry", "<Button-3>", self._show_entry_popup)

    def _show_entry_popup(self, event):
        self.current_entry = event.widget
        self._popup_selection_range = None
        if self.current_entry.select_present():
            self._popup_selection_range = (self.current_entry.index("sel.first"), self.current_entry.index("sel.last"))
        self.entry_popup.tk_popup(event.x_root, event.y_root)

    def _copy_entry(self):
        try:
            if getattr(self, 'current_entry', None):
                if self.current_entry.select_present():
                    first, last = self.current_entry.index("sel.first"), self.current_entry.index("sel.last")
                elif getattr(self, '_popup_selection_range', None):
                    first, last = self._popup_selection_range
                else:
                    return
                text = self.current_entry.get()[first:last]
                self.root.clipboard_clear()
                self.root.clipboard_append(text)
                self.root.update()
        except Exception: pass

    def _paste_entry(self):
        try:
            text = self.root.clipboard_get()
            if getattr(self, 'current_entry', None):
                if getattr(self, '_popup_selection_range', None):
                    first, last = self._popup_selection_range
                    self.current_entry.delete(first, last)
                    self.current_entry.insert(first, text)
                elif self.current_entry.select_present():
                    self.current_entry.delete("sel.first", "sel.last")
                    self.current_entry.insert(tk.INSERT, text)
                else:
                    self.current_entry.insert(tk.INSERT, text)
        except Exception: pass

    def _cut_entry(self):
        self._copy_entry()
        try:
            if getattr(self, 'current_entry', None):
                if getattr(self, '_popup_selection_range', None):
                    first, last = self._popup_selection_range
                    self.current_entry.delete(first, last)
                elif self.current_entry.select_present():
                    self.current_entry.delete("sel.first", "sel.last")
        except Exception: pass

    def _refresh_treeview(self):
        """Re-render the treeview with updated columns layout"""
        self.column_keys = [c[0] for c in self.columns_config]
        self.display_columns = [c[1] for c in self.columns_config]
        self.column_widths = {c[1]: c[2] for c in self.columns_config}
        
        self.preview_tree['columns'] = self.display_columns
        for col, width in self.column_widths.items():
            self.preview_tree.heading(col, text=col)
            self.preview_tree.column(col, width=width, anchor='center')
        
        # Repopulate
        self.populate_preview(self.extracted_data, switch_tab=False)
        self.save_settings()

    def open_column_config_window(self):
        """Open a window to configure treeview columns"""
        win = tk.Toplevel(self.root)
        win.title(f"{self.tab_name} 컬럼 설정")
        win.geometry("500x400")
        win.configure(background="#f9fafb")
        win.grab_set()

        ttk.Label(win, text="미리보기 컬럼 설정", font=("Malgun Gothic", 12, "bold"), background="#f9fafb").pack(pady=10)
        
        list_frame = tk.Frame(win, background="#f9fafb")
        list_frame.pack(fill='both', expand=True, padx=20, pady=5)
        
        # columns listbox
        cols_lb = tk.Listbox(list_frame, font=("Malgun Gothic", 10), selectmode=tk.SINGLE)
        cols_lb.pack(side='left', fill='both', expand=True)
        
        scr = ttk.Scrollbar(list_frame, orient="vertical", command=cols_lb.yview)
        cols_lb.configure(yscrollcommand=scr.set)
        scr.pack(side='left', fill='y')
        
        def refresh_list():
            cols_lb.delete(0, tk.END)
            for c in self.columns_config:
                flags = "(기본, 삭제불가)" if c[0] in ["selected", "No", "Date", "Dwg", "Joint", "Loc"] else ""
                cols_lb.insert(tk.END, f"{c[1]} [{c[0]}] {flags}")

        refresh_list()
        
        # Right buttons
        btn_frame = tk.Frame(list_frame, background="#f9fafb", padx=10)
        btn_frame.pack(side='right', fill='y')
        
        def add_col():
            key = simpledialog.askstring("컬럼 추가", "데이터 키(Key) 이름을 입력하세요 (영문 권장):", parent=win)
            if not key: return
            disp = simpledialog.askstring("컬럼 추가", "표시될 컬럼 이름을 입력하세요:", parent=win)
            if not disp: return
            width = simpledialog.askinteger("컬럼 넓이", "컬럼 넓이를 입력하세요 (보통 80~100):", initialvalue=80, parent=win)
            if not width: width = 80
            self.columns_config.append((key, disp, width))
            refresh_list()
            
        def del_col():
            sel = cols_lb.curselection()
            if not sel: return
            idx = sel[0]
            if self.columns_config[idx][0] in ["selected", "No", "Date", "Dwg", "Joint", "Loc"]:
                messagebox.showwarning("삭제 불가", "이 컬럼은 시스템 필수 컬럼으로 삭제할 수 없습니다.", parent=win)
                return
            if messagebox.askyesno("삭제", f"'{self.columns_config[idx][1]}' 컬럼을 삭제하시겠습니까?", parent=win):
                del self.columns_config[idx]
                refresh_list()
                
        def rename_col():
            sel = cols_lb.curselection()
            if not sel: return
            idx = sel[0]
            new_disp = simpledialog.askstring("이름 변경", "새로운 표시 이름을 입력하세요:", initialvalue=self.columns_config[idx][1], parent=win)
            if new_disp:
                c = self.columns_config[idx]
                self.columns_config[idx] = (c[0], new_disp, c[2])
                refresh_list()

        def move_up():
            sel = cols_lb.curselection()
            if not sel: return
            idx = sel[0]
            if idx > 0:
                self.columns_config[idx], self.columns_config[idx-1] = self.columns_config[idx-1], self.columns_config[idx]
                refresh_list()
                cols_lb.selection_set(idx-1)

        def move_down():
            sel = cols_lb.curselection()
            if not sel: return
            idx = sel[0]
            if idx < len(self.columns_config) - 1:
                self.columns_config[idx], self.columns_config[idx+1] = self.columns_config[idx+1], self.columns_config[idx]
                refresh_list()
                cols_lb.selection_set(idx+1)

        ttk.Button(btn_frame, text="추가", command=add_col).pack(fill='x', pady=2)
        ttk.Button(btn_frame, text="이름 변경", command=rename_col).pack(fill='x', pady=2)
        ttk.Button(btn_frame, text="삭제", command=del_col).pack(fill='x', pady=2)
        ttk.Label(btn_frame, text=" ", background="#f9fafb").pack(pady=5)
        ttk.Button(btn_frame, text="▲ 위로", command=move_up).pack(fill='x', pady=2)
        ttk.Button(btn_frame, text="▼ 아래로", command=move_down).pack(fill='x', pady=2)

        def apply_and_close():
            self._refresh_treeview()
            win.destroy()

        ttk.Button(win, text="적용 및 닫기", style="Action.TButton", command=apply_and_close).pack(pady=15)

    def on_tree_select(self, event):
        """React to selection changes without blocking mouse drag-selection."""
        pass

    def on_tree_double_click(self, event):
        """Dynamic column double click for inline editing"""
        region = self.preview_tree.identify_region(event.x, event.y)
        if region != "cell": return
        
        column = self.preview_tree.identify_column(event.x)
        col_idx = int(column.replace("#", "")) - 1
        
        if col_idx < 0 or col_idx >= len(self.column_keys): return
        
        # 'selected' 컬럼 등 편집 불가 컬럼 필터링
        key = self.column_keys[col_idx]
        if key in ["selected", "No", "Date"]: return
        
        item_id = self.preview_tree.identify_row(event.y)
        if not item_id: return
        
        x, y, w, h = self.preview_tree.bbox(item_id, column)
        
        view_idx = self.preview_tree.index(item_id)
        if hasattr(self, 'item_idx_map') and 0 <= view_idx < len(self.item_idx_map):
            actual_idx = self.item_idx_map[view_idx]
            old_val = self.extracted_data[actual_idx].get(key, "")
        else:
            old_val = self.preview_tree.set(item_id, column)
        
        entry = ttk.Entry(self.preview_tree, exportselection=False)
        entry.insert(0, old_val)
        entry.select_range(0, tk.END)
        entry.place(x=x, y=y, width=w, height=h)
        entry.focus_set()
        
        def finish_edit(event=None):
            new_val = entry.get()
            view_idx = self.preview_tree.index(item_id)
            if hasattr(self, 'item_idx_map') and 0 <= view_idx < len(self.item_idx_map):
                actual_idx = self.item_idx_map[view_idx]
                self.extracted_data[actual_idx][key] = new_val
                self.populate_preview(self.extracted_data, switch_tab=False)
            entry.destroy()
            
        entry.bind("<Return>", finish_edit)
        entry.bind("<FocusOut>", finish_edit)
        entry.bind("<Escape>", lambda e: entry.destroy())

    def show_context_menu(self, event):
        """우클릭 컨텍스트 메뉴 표시 및 마지막 클릭 컬럼 기록"""
        col = self.preview_tree.identify_column(event.x)
        self.last_clicked_col = col # Track for copy/paste
        # [NEW] 우클릭 시 드래그 범위를 해당 칸으로 초기화하여 엉뚱한 칸에 붙여넣기 방지
        self.start_col = col
        self.end_col = col
        
        item_id = self.preview_tree.identify_row(event.y)
        if item_id:
            # 클릭한 항목이 선택되어 있지 않으면 해당 항목만 선택
            if item_id not in self.preview_tree.selection():
                self.preview_tree.selection_set(item_id)
        # 항목이 없어도 메뉴는 띄워줌 (붙여넣기 등을 위해)
        self.ctx_menu.post(event.x_root, event.y_root)


    def update_date_listbox(self):
        if not hasattr(self, 'date_listbox'): return
        unique_dates = sorted(list(set(item.get('Date', '') for item in self.extracted_data if item.get('Date'))))
        self.date_listbox.delete(0, tk.END)
        for i, d in enumerate(unique_dates):
            self.date_listbox.insert(tk.END, f"[v] {d}")

    def add_manual_row(self, show_msg=True, refresh=True):
        """데이터 추출 파일 없이 수동으로 빈 행 추가 (선택된 행 바로 아래에 추가)"""
        new_row = {k: "" for k in self.column_keys}
        new_row['selected'] = True
        
        # 기본 날짜를 오늘로 설정 (또는 기존 데이터의 마지막 날짜)
        last_date = ""
        if self.extracted_data:
            last_date = self.extracted_data[-1].get('Date', '')
        
        new_row['Date'] = last_date if last_date else datetime.datetime.now().strftime("%Y-%m-%d")
        
        # 순번(No) 자동 계산
        if self.extracted_data:
            try:
                # 숫자형 NO 중 최대값 + 1
                nums = [int(str(x.get('No', '0'))) for x in self.extracted_data if str(x.get('No', '')).isdigit()]
                new_row['No'] = max(nums) + 1 if nums else len(self.extracted_data) + 1
            except:
                new_row['No'] = len(self.extracted_data) + 1
        else:
            new_row['No'] = 1
            
        # [MOD] 현재 선택된 위치 아래에 삽입 (없으면 맨 뒤)
        selected_items = self.preview_tree.selection()
        insertion_index = len(self.extracted_data)
        
        if selected_items:
            # 마지막으로 선택된 항목 이후에 삽입
            last_selected_id = selected_items[-1]
            try:
                # Treeview 상의 실제 인덱스 (필터링된 상태 고려)
                idx = self.preview_tree.index(last_selected_id)
                insertion_index = idx + 1
            except: pass

        self.extracted_data.insert(insertion_index, new_row)
        
        if refresh:
            self.populate_preview(self.extracted_data, switch_tab=False)
            self.update_date_listbox()
            
            # [NEW] 새로 추가된 행 자동으로 선택 및 포커스
            try:
                children = self.preview_tree.get_children()
                if insertion_index < len(children):
                    new_item_id = children[insertion_index]
                    self.preview_tree.selection_set(new_item_id)
                    self.preview_tree.see(new_item_id)
                    self.preview_tree.focus(new_item_id)
            except: pass
            
            if show_msg:
                self.log("➕ 선택된 행 아래에 새 행이 추가되었습니다.")
        return new_row

    def copy_cell(self):
        """선택된 영역(드래그된 컬럼 범위)의 내용을 클립보드에 복사 (엑셀 스마트 영역 복사)"""
        selected = self.preview_tree.selection()
        if not selected: return
        
        # [SMART AREA] 드래그 시작열과 종료열 사이의 범위를 계산
        try:
            s_idx = int(str(getattr(self, 'start_col', '#4')).replace('#', '')) - 1
            e_idx = int(str(getattr(self, 'end_col', '#4')).replace('#', '')) - 1
            
            # [ENHANCED] 드래그 범위가 없을 때(그냥 행 클릭)는 '선택/No' 제외한 모든 의미있는 열을 자동으로 범위로 잡음
            if s_idx == e_idx:
                c_start = 2 # 'Date' 컬럼 (#3)
                c_end = len(self.column_keys) - 1
            else:
                c_start = min(s_idx, e_idx)
                c_end = max(s_idx, e_idx)
        except:
            c_start = 2 # Default to all data columns
            c_end = len(self.column_keys) - 1
            
        target_col_ids = [f"#{i+1}" for i in range(c_start, c_end + 1)]
        
        copied_rows = []
        for item_id in selected:
            row_vals = []
            for c_id in target_col_ids:
                val = self.preview_tree.set(item_id, c_id)
                row_vals.append(str(val))
            copied_rows.append("\t".join(row_vals)) # Use TAB for Excel columns

        final_string = "\n".join(copied_rows)
        self.root.clipboard_clear()
        self.root.clipboard_append(final_string)
        self.root.update()
        self.log(f"📋 엑셀 영역 복사 완료: {len(selected)}행 x {len(target_col_ids)}열")

    def duplicate_selected_rows(self):
        """선택된 항목들을 그대로 복제하여 맨 아래에 추가 (데이터 모델 우선)"""
        selected_items = list(self.preview_tree.selection())
        selected_items.sort(key=lambda x: self.preview_tree.index(x))
        if not selected_items: return
        
        duplicated_count = 0
        for itm in selected_items:
            v_idx = self.preview_tree.index(itm)
            if hasattr(self, 'item_idx_map') and 0 <= v_idx < len(self.item_idx_map):
                original_idx = self.item_idx_map[v_idx]
                original_data = self.extracted_data[original_idx]
                
                # 데이터 깊은 복사 (ID, No 등은 제외하거나 새로 생성)
                new_row = original_data.copy()
                
                # 순번(No) 재계산 (max + 1)
                try:
                    nums = [int(str(x.get('No', '0'))) for x in self.extracted_data if str(x.get('No', '')).isdigit()]
                    new_row['No'] = max(nums) + 1 if nums else len(self.extracted_data) + 1
                except:
                    new_row['No'] = len(self.extracted_data) + 1
                
                self.extracted_data.append(new_row)
                duplicated_count += 1
        
        # UI 리프레시
        self.populate_preview(self.extracted_data, switch_tab=False)
        self.log(f"👯 선택된 {duplicated_count}개 행을 복제하여 추가했습니다.")

    def paste_cell(self):
        """클립보드 내용을 데이터 모델에 직접 업데이트 후 UI 리프레시 (ID 유실 문제 해결)"""
        try: 
            clipboard_val = self.root.clipboard_get()
            if not clipboard_val: return
        except: return
        
        # 1. 클립보드 파싱
        raw_rows = [line.split("\t") for line in clipboard_val.splitlines()]
        # 마지막 빈 줄 제거 (엑셀 복사 시 흔함)
        if raw_rows and len(raw_rows) > 1 and not any(raw_rows[-1]):
            raw_rows.pop()
        if not raw_rows: return

        # 2. 대상 데이터 인덱스 확인
        selected_items = list(self.preview_tree.selection())
        selected_items.sort(key=lambda x: self.preview_tree.index(x))
        
        target_indices = []
        for itm in selected_items:
            v_idx = self.preview_tree.index(itm)
            if hasattr(self, 'item_idx_map') and 0 <= v_idx < len(self.item_idx_map):
                target_indices.append(self.item_idx_map[v_idx])

        # 선택된 게 없으면 새 행부터 시작
        if not target_indices:
            target_indices = [len(self.extracted_data)]

        # 3. 붙여넣을 컬럼 범위 계산
        try:
            s_col = getattr(self, 'start_col', getattr(self, 'last_clicked_col', '#4'))
            e_col = getattr(self, 'end_col', s_col)
            s_idx = int(str(s_col).replace('#', '')) - 1
            e_idx = int(str(e_col).replace('#', '')) - 1
            is_explicit_range = (s_idx != e_idx)
            
            c_start = min(s_idx, e_idx)
            c_end = max(s_idx, e_idx)
            
            # 단일 클릭 시 클립보드 실제 데이터 폭만큼 확장 (단, trailing empty tabs는 유지)
            if not is_explicit_range:
                clip_width = len(raw_rows[0])
                if clip_width > 1:
                    c_end = min(c_start + clip_width - 1, len(self.column_keys) - 1)
        except:
            c_start = c_end = 3
            is_explicit_range = False
            
        target_col_keys = self.column_keys[c_start : c_end + 1]

        # 4. Fill Down / Multi-row 행 확장
        if len(target_indices) == 1 and len(raw_rows) > 1:
            base_idx = target_indices[0]
            target_indices = [base_idx + i for i in range(len(raw_rows))]

        # 5. 필요한 만큼 데이터 행 미리 생성 (UI 리프레시 없이)
        max_needed = max(target_indices)
        while len(self.extracted_data) <= max_needed:
            self.add_manual_row(show_msg=False, refresh=False)

        # 6. 데이터 모델 직접 수정
        for r_offset, d_idx in enumerate(target_indices):
            # 붙여넣을 데이터 행 (부족하면 마지막 행 반복)
            source_row = raw_rows[min(r_offset, len(raw_rows)-1)]
            
            for c_offset, key in enumerate(target_col_keys):
                # 단일 클릭 붙여넣기 시 클립보드 범위를 넘어가면 중단
                if c_offset >= len(source_row) and not is_explicit_range:
                    break
                
                # 값 가져오기 (컬럼 범위가 데이터보다 넓으면 마지막 값 반복)
                val = source_row[min(c_offset, len(source_row)-1)].strip()
                
                if key not in ["selected", "No"]:
                    self.extracted_data[d_idx][key] = val

        # 7. UI 전체 리프레시 (단 한 번)
        self.populate_preview(self.extracted_data, switch_tab=False)
        self.preview_tree.focus_set()
        self.log(f"📋 붙여넣기 완료: {len(target_indices)}행 x {len(target_col_keys)}열")

    def merge_selected_iso(self):
        """선택된 항목들의 ISO 번호를 첫 번째 항목의 것으로 통일 (병합 효과)"""
        selected = self.preview_tree.selection()
        if len(selected) < 2:
            messagebox.showinfo("알림", "병합할 항목을 2개 이상 선택해주세요.")
            return
        
        first_iso = self.preview_tree.set(selected[0], "#4")
        for item_id in selected:
            view_idx = self.preview_tree.index(item_id)
            if hasattr(self, 'item_idx_map') and 0 <= view_idx < len(self.item_idx_map):
                actual_idx = self.item_idx_map[view_idx]
                self.extracted_data[actual_idx]['Dwg'] = first_iso
        
        self.populate_preview(self.extracted_data, switch_tab=False)
        self.log(f"🔗 {len(selected)}개 항목 ISO 병합 완료: {first_iso}")
        messagebox.showinfo("성공", f"{len(selected)}개 항목의 ISO 번호를 '{first_iso}'(으)로 병합했습니다.")

    def merge_selected_joint(self):
        """선택된 항목들의 Joint No를 첫 번째 항목의 것으로 통일 (병합 효과)"""
        selected = self.preview_tree.selection()
        if len(selected) < 2:
            messagebox.showinfo("알림", "병합할 항목을 2개 이상 선택해주세요.")
            return
        
        # 첫 번째 아이템의 Joint No 추출
        item_id = selected[0]
        view_idx = self.preview_tree.index(item_id)
        if hasattr(self, 'item_idx_map') and 0 <= view_idx < len(self.item_idx_map):
            actual_idx = self.item_idx_map[view_idx]
            first_joint = self.extracted_data[actual_idx].get('Joint', '')
        else:
            first_joint = self.preview_tree.set(item_id, "#5")

        for item_id in selected:
            view_idx = self.preview_tree.index(item_id)
            if hasattr(self, 'item_idx_map') and 0 <= view_idx < len(self.item_idx_map):
                actual_idx = self.item_idx_map[view_idx]
                self.extracted_data[actual_idx]['Joint'] = first_joint
        
        self.populate_preview(self.extracted_data, switch_tab=False)
        self.log(f"🔗 {len(selected)}개 항목 Joint 병합 완료: {first_joint}")
        messagebox.showinfo("성공", f"{len(selected)}개 항목의 Joint No를 '{first_joint}'(으)로 병합했습니다.")

    def select_all(self):
        """모든 항목 체크"""
        for item in self.extracted_data: item['selected'] = True
        self.populate_preview(self.extracted_data, switch_tab=False)

    def deselect_all(self):
        """모든 항목 체크 해제"""
        for item in self.extracted_data: item['selected'] = False
        self.populate_preview(self.extracted_data, switch_tab=False)

    def move_item(self, direction):
        """선택된 아이템의 순서를 위/아래로 이동 (필터 상태 대응)"""
        selected_ids = self.preview_tree.selection()
        if not selected_ids: return
        
        # 1. 보기 인덱스 확인
        view_indices = sorted([self.preview_tree.index(sid) for sid in selected_ids])
        
        # 2. 경계 검사
        if direction == -1: # 위로
            if view_indices[0] == 0: return
            processing_order = view_indices
        else: # 아래로
            if view_indices[-1] == len(self.item_idx_map) - 1: return
            processing_order = reversed(view_indices)
            
        # 3. item_idx_map을 이용해 실제 데이터 리스트(extracted_data)에서 스왑
        for v_idx in processing_order:
            actual_curr = self.item_idx_map[v_idx]
            actual_target = self.item_idx_map[v_idx + direction]
            
            self.extracted_data[actual_curr], self.extracted_data[actual_target] = \
                self.extracted_data[actual_target], self.extracted_data[actual_curr]
        
        # 4. 보기 갱신
        self.populate_preview(self.extracted_data, switch_tab=False)
        
        # 5. 선택 상태 복구
        new_items = self.preview_tree.get_children()
        for v_idx in view_indices:
            self.preview_tree.selection_add(new_items[v_idx + direction])

    def delete_item(self):
        """선택된 아이템 삭제"""
        selected_items = self.preview_tree.selection()
        if not selected_items: return
        if not messagebox.askyesno("삭제 확인", f"선택한 {len(selected_items)}개 항목을 삭제하시겠습니까?"): return
        
        indices = sorted([self.item_idx_map[self.preview_tree.index(s)] for s in selected_items], reverse=True)
        for idx in indices:
            self.extracted_data.pop(idx)
        self.populate_preview(self.extracted_data, switch_tab=False)
        self.update_date_listbox()

    def clear_all(self):
        """추출 목록 초기화"""
        if not self.extracted_data: return
        if not messagebox.askyesno("초기화 확인", "누적된 모든 데이터를 삭제하고 초기화하시겠습니까?"): return
        self.extracted_data = []
        self.populate_preview(self.extracted_data, switch_tab=False)
        self.update_date_listbox()
        self.log("🧹 데이터 목록이 초기화되었습니다.")

    def populate_preview(self, data_list, switch_tab=True):
        """추출된 데이터를 미리보기 표에 채움 (필터 반영 및 그룹 색상 적용)"""
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        
        self.item_idx_map = []
        filter_enabled = self.show_selected_only.get()
        
        last_iso = None
        last_joint = None
        current_tag = "group_even"
        
        for idx, item in enumerate(data_list):
            is_selected = item.get('selected', True)
            if filter_enabled and not is_selected:
                continue
            
            self.item_idx_map.append(idx)
            v_mark = "●" if is_selected else "○"
            
            # ISO/DWG 번호가 바뀌면 배경색 태그 교체
            curr_iso = item.get('Dwg', '')
            curr_joint = item.get('Joint', '')
            display_iso = curr_iso
            display_joint = curr_joint

            if last_iso is not None and curr_iso != last_iso:
                current_tag = "group_odd" if current_tag == "group_even" else "group_even"
                # ISO가 바뀌면 Joint도 표시 (last_joint 초기화)
                last_joint = None 
            elif last_iso is not None and curr_iso == last_iso:
                # 이전 행과 ISO가 동일하면 화면 표시 생략 (병합 효과)
                display_iso = ""
                # ISO가 같을 때 Joint도 같으면 Joint 표시 생략
                if curr_joint == last_joint:
                    display_joint = ""
                
            last_iso = curr_iso
            last_joint = curr_joint
            
            # 동적 컬럼 값 매핑
            row_values = []
            for key in self.column_keys:
                if key == "selected":
                    row_values.append(v_mark)
                elif key == "Dwg":
                    row_values.append(display_iso)
                elif key == "Joint":
                    row_values.append(display_joint)
                else:
                    row_values.append(item.get(key, ''))
            
            self.preview_tree.insert("", "end", values=tuple(row_values), tags=(str(idx), current_tag))
            
        if switch_tab:
            self.tab_notebook.select(self.tab_preview)

    def _browse_dir(self, var):
        path = filedialog.askdirectory(initialdir=var.get() or SCRIPT_HOME)
        if path: var.set(path)

    def _browse_file(self, var, types):
        path = filedialog.askopenfilename(initialdir=os.path.dirname(var.get() or SCRIPT_HOME), filetypes=types)
        if path: var.set(path)

    # --- Integrated Verification Logic ---

    def to_float(self, val):
        if pd.isna(val): return 0.0
        s = str(val).upper().replace("%", "").strip()
        if "<" in s or "ND" in s or s == "": return 0.0
        try: return float(s)
        except: return 0.0

    def check_material_grade(self, row_data):
        """jjchRFIPMI.py에서 이식된 10% 여유치 판정 로직"""
        cr = row_data.get('Cr', 0.0)
        ni = row_data.get('Ni', 0.0)
        mo = row_data.get('Mo', 0.0)
        mn = row_data.get('Mn', 0.0)
        
        margin = 0.1 # 10% 여유
        
        # 1. SUS 316 (Cr:16~18 / Ni:10~14 / Mo:2~3)
        if (16.0*(1-margin) <= cr <= 18.0*(1+margin)) and (10.0*(1-margin) <= ni <= 14.0*(1+margin)) and (2.0*(1-margin) <= mo <= 3.0*(1+margin)):
            return "SS316"
        
        # 2. Duplex (Cr:22~23 / Mo:3~3.5 / Ni:4.5~6.5 / Mn:2.0이하)
        if (22.0*(1-margin) <= cr <= 23.0*(1+margin)) and (4.5*(1-margin) <= ni <= 6.5*(1+margin)) and (3.0*(1-margin) <= mo <= 3.5*(1+margin)) and (mn <= 2.2):
            return "DUPLEX"

        # 3. SUS 310 (Cr:24~26 / Ni:19~22)
        if (24.0*(1-margin) <= cr <= 26.0*(1+margin)) and (19.0*(1-margin) <= ni <= 22.0*(1+margin)):
            return "SS310"

        # 4. SUS 304 (Cr:18↑ / Ni:8↑ / Mo:0.5↓ / Mn:2.0↓)
        if (cr >= 16.2) and (ni >= 7.2) and (mo <= 0.55) and (mn <= 2.2):
            return "SS304"

        return None # 판정 불가 시 원본 사용

    # --- Excel Helper Logic ---

    def find_image_smart(self, keyword, exclude_keyword=None):
        def _search_in_folder(folder_path):
            if not folder_path or not os.path.exists(folder_path): return None
            candidates = glob.glob(os.path.join(folder_path, "*.*"))
            valid_extensions = ['.PNG', '.JPG', '.JPEG', '.BMP', '.GIF']
            for path in candidates:
                fname = os.path.basename(path).upper(); ext = os.path.splitext(path)[1].upper()
                if ext not in valid_extensions: continue 
                if keyword.upper() in fname:
                    if exclude_keyword and exclude_keyword.upper() in fname: continue
                    return path 
            return None

        # 1. UI에서 설정한 폴더에서 먼저 검색
        found = _search_in_folder(self.logo_folder_path.get())
        if found: return found
        
        # 2. PyInstaller 묶음(실행파일 내부 임시 폴더)에서 검색 (Standalone 지원)
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            found = _search_in_folder(sys._MEIPASS)
            if found: return found
            
        return None

    def place_image_freely(self, ws, img_path, anchor_cell_str, w, h, x_offset, y_offset):
        try:
            if not img_path or not os.path.exists(img_path): return
            original = PILImage.open(img_path).convert("RGBA")
            resized = original.resize((int(w), int(h)), PILImage.Resampling.LANCZOS)
            temp_name = f"temp_{os.path.basename(img_path)}"
            temp_full_path = os.path.join(tempfile.gettempdir(), temp_name)
            resized.save(temp_full_path)
            
            img = XLImage(temp_full_path); img.width = w; img.height = h
            col_str, row_num = coordinate_from_string(anchor_cell_str)
            col_idx = column_index_from_string(col_str) - 1; row_idx = row_num - 1 
            emu_x = int(x_offset * 9525); emu_y = int(y_offset * 9525)
            emu_w = int(w * 9525); emu_h = int(h * 9525)
            marker = AnchorMarker(col=col_idx, colOff=emu_x, row=row_idx, rowOff=emu_y)
            size = XDRPositiveSize2D(cx=emu_w, cy=emu_h)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(img)
        except Exception as e: self.log(f"[WARNING] {os.path.basename(img_path)} 배치 실패: {e}")

    def add_logos_to_sheet(self, ws, is_cover=False, clear_existing=True):
        # [NEW] RT 성적서의 경우 템플릿에 Shooting Sketch나 로고가 이미 포함되어 있는 경우가 많습니다.
        # openpyxl 라이브러리의 한계로 인해, 기존 그림이 있는 시트에 add_image()를 호출하면 
        # 기존 그림이 사라지는 현상이 발생하므로 RT 탭에서는 로고 추가를 스킵합니다.
        # [MOD] RT 성적서의 경우, 이미 이미지가 있다면 템플릿 보호를 위해 스킵 (openpyxl 한계 방어)
        if self.tab_name == "RT":
            if hasattr(ws, '_images') and len(ws._images) > 0:
                return

        if clear_existing:
            # PMI 등에서만 명시적으로 기존 이미지를 비움. RT 등은 템플릿 보호를 위해 비우지 않음.
            if self.tab_name == "PMI":
                try: ws._images = [] 
                except: pass
        
        ctx = "COVER" if is_cover else "DATA"

        # [FIX] RT 성적서의 경우, 이미 템플릿에 로고가 크게 박혀있을 수 있으므로 
        # 로고가 겹쳐서 나오지 않도록 PMI 로고(stamp 등) 필터링 강화
        exclude = "PMI" if self.tab_name != "PMI" else "PMI-1"

        def _get_params(key_prefix):
            return (self.config[f"{key_prefix}_{ctx}_ANCHOR"], self.config[f"{key_prefix}_{ctx}_W"], self.config[f"{key_prefix}_{ctx}_H"], self.config[f"{key_prefix}_{ctx}_X"], self.config[f"{key_prefix}_{ctx}_Y"])
        
        # SITCO/서울검사 로고 (PMI 키워드 배제)
        path = self.config.get(f"SITCO_{ctx}_PATH")
        if not path or not os.path.exists(path):
            path = self.find_image_smart("SITCO", exclude_keyword=exclude)
        self.place_image_freely(ws, path, *_get_params("SITCO"))
        
        path = self.config.get(f"SEOUL_{ctx}_PATH")
        if not path or not os.path.exists(path):
            path = self.find_image_smart("서울검사", exclude_keyword=exclude)
        self.place_image_freely(ws, path, *_get_params("SEOUL"))
        
        # 바닥글 (Footer): 탭별 전용 바닥글(예: RT바닥글)을 먼저 찾고, 없으면 공용 바닥글(PMI 제외)을 찾음
        path = self.config.get(f"FOOTER_{ctx}_PATH")
        if not path or not os.path.exists(path):
            path = self.find_image_smart(f"{self.tab_name}바닥글")
            if not path:
                path = self.find_image_smart("바닥글", exclude_keyword=exclude)
                if not path and self.tab_name == "PMI": path = self.find_image_smart("PMI", exclude_keyword="PMI-1")
        self.place_image_freely(ws, path, *_get_params("FOOTER"))
        
        # 왼쪽 하단 로고/스탬프: 탭 이름(RT 등)으로 먼저 찾고, PMI일 때만 PMI-1 등을 허용
        path_left = self.config.get(f"FOOTER_PT_{ctx}_PATH")
        if not path_left or not os.path.exists(path_left):
            if is_cover and self.tab_name == "RT":
                path_left = self.find_image_smart("바닥글 좌측 갑")
            if not path_left:
                path_left = self.find_image_smart(f"{self.tab_name}갑") if is_cover else None
            if not path_left:
                path_left = self.find_image_smart(self.tab_name, exclude_keyword=None if self.tab_name == "PMI" else "PMI")
            if not path_left and self.tab_name == "PMI":
                path_left = self.find_image_smart("PMI-1")
            
            
        self.place_image_freely(ws, path_left, *_get_params("FOOTER_PT"))

        # [NEW] Shooting Sketch (RT 갑지만 해당) - 상/하 분리
        if self.tab_name == "RT" and is_cover:
            # 상단 스케치
            path_up = self.config.get(f"SKETCH_UP_{ctx}_PATH")
            if not path_up or not os.path.exists(path_up):
                path_up = (self.find_image_smart("Shooting Sketch 상") or 
                           self.find_image_smart("Shooting Sketch-1") or 
                           self.find_image_smart("Shooting Sketch 1") or
                           self.find_image_smart("Shooting Sketch1"))
            self.log(f"📸 Shooting Sketch (상) 경로: {path_up}")
            self.place_image_freely(ws, path_up, *_get_params("SKETCH_UP"))

            # 하단 스케치
            path_down = self.config.get(f"SKETCH_DOWN_{ctx}_PATH")
            if not path_down or not os.path.exists(path_down):
                path_down = (self.find_image_smart("Shooting Sketch 하") or 
                             self.find_image_smart("Shooting Sketch-2") or 
                             self.find_image_smart("Shooting Sketch 2") or
                             self.find_image_smart("Shooting Sketch2"))
            self.log(f"📸 Shooting Sketch (하) 경로: {path_down}")
            self.place_image_freely(ws, path_down, *_get_params("SKETCH_DOWN"))
            
            if not path_up and not path_down:
                self.log("⚠️ Shooting Sketch 이미지를 찾을 수 없습니다. (상/하 키워드 확인 필요)")

    def force_print_settings(self, ws):
        try:
            # [MOD] RT 탭의 을시트 인쇄 영역은 A1:X40으로 설정 (사용자 요청)
            if self.tab_name == "RT":
                ws.print_area = 'A1:X40'
            else:
                ws.print_area = f'A1:M{self.config["PRINT_END_ROW"]}'
            
            ws.page_setup.paperSize = 9; ws.page_setup.orientation = 'portrait'; ws.page_setup.scale = 95
            ws.print_options.horizontalCentered = True; ws.print_options.verticalCentered = True
            ws.page_margins.left = 0.5; ws.page_margins.right = 0.3; ws.page_margins.top = 0.2; ws.page_margins.bottom = 0.2
        except: pass

    def set_eulji_headers(self, ws):
        if self.tab_name != "PMI": return # [RT/NDT] 템플릿의 고유 헤더 보호
        headers = ["NI", "CR", "MO"]
        data_font = Font(size=9); header_row = self.config['START_ROW']
        for c in range(7, 14): ws.cell(row=header_row, column=c).value = None
        for i, val in enumerate(headers):
            col = 8 + i
            cell = ws.cell(row=header_row, column=col)
            cell.value = val; cell.alignment = Alignment(horizontal='center', vertical='center'); cell.font = data_font
        
        materials = "SS304,SS304L,SS316,SS316L,SS321,SS347,SS410,SS430,DUPLEX,MONEL,INCONEL,ER308,ER308L,ER309,ER309L,ER316,ER316L,ER347,ER2209,WP316,WP316L,TP316,TP316L,F316L,A182-F316L,A312-TP316L"
        dv_q = DataValidation(type="list", formula1=f'"{materials}"', allow_blank=True)
        ws.add_data_validation(dv_q)
        for r in range(self.config['START_ROW'], self.config['DATA_END_ROW'] + 1):
            target_l = ws.cell(row=r, column=13); dv_q.add(target_l)
            target_l.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center'); target_l.font = Font(size=8.5)

    def safe_set_value(self, ws, coord, value, align=None):
        """Safely set value to a cell even if it is part of a merged range."""
        try:
            cell = ws[coord]
            target_cell = cell
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                for m_range in ws.merged_cells.ranges:
                    if coord in m_range:
                        target_cell = ws.cell(row=m_range.min_row, column=m_range.min_col)
                        break
            
            target_cell.value = value
            if align:
                target_cell.alignment = Alignment(horizontal=align, vertical='center')
        except Exception as e:
            self.log(f"⚠️ Cell write failed at {coord}: {e}")

    def prepare_next_sheet(self, wb, source_sheet_idx, page_num, total_pages=1):
        source_sheet = wb.worksheets[source_sheet_idx]; new_sheet = wb.copy_worksheet(source_sheet) 
        base_title = source_sheet.title.split('_')[0]; new_sheet.title = f"{base_title[:20]}_{page_num:03d}"
        self.force_print_settings(new_sheet); self.add_logos_to_sheet(new_sheet, is_cover=False, clear_existing=False)
        for col_letter, col_dim in source_sheet.column_dimensions.items(): new_sheet.column_dimensions[col_letter].width = col_dim.width
        data_font = Font(size=9); grade_font = Font(size=8.5)
        for r in range(self.config['START_ROW'], self.config['DATA_END_ROW'] + 1):
            if self.tab_name == "PMI":
                rd = new_sheet.row_dimensions[r]; rd.height = 20.55
            for c in range(1, 14):
                cell = new_sheet.cell(row=r, column=c)
                cell.font = grade_font if c == 13 else data_font
                if not isinstance(cell, MergedCell): cell.value = None
        merged_to_clear = [rng for rng in new_sheet.merged_cells.ranges if rng.min_row >= self.config['START_ROW'] and rng.max_row <= self.config['DATA_END_ROW']]
        for rng in merged_to_clear: new_sheet.unmerge_cells(str(rng))
        
        # [NEW] 이미지 및 레거시 드로잉(바닥글 등) 수동 복사 (RT 제외)
        if self.tab_name != "RT":
            if hasattr(source_sheet, '_images'):
                from copy import deepcopy
                for img in source_sheet._images:
                    new_img = deepcopy(img)
                    new_sheet.add_image(new_img)
            
            # 레거시 드로잉(바닥글 스케치 등) 보존 시도
            if hasattr(source_sheet, 'legacy_drawing'):
                new_sheet.legacy_drawing = source_sheet.legacy_drawing
        else:
            # [RT 전용] 이미지 복사 대신 로고 재삽입 (데이터 시트용)
            # 을지에 스케치가 이미 포함된 템플릿이 아님을 확인한 후 삽입함
            self.add_logos_to_sheet(new_sheet, is_cover=False, clear_existing=False)

        # 페이지 번호 업데이트 (Page N of M)
        if self.tab_name == "RT":
            # [MOD] RT 성적서 을지 페이지 번호 형식 (Page X of Y)
            # R2와 T2가 병합되어 있을 가능성을 고려하여 R2에만 값을 넣음
            self.safe_set_value(new_sheet, 'R2', f"Page   {page_num}  of   {total_pages}")
        else:
            self.safe_set_value(new_sheet, 'R2', page_num)
            self.safe_set_value(new_sheet, 'T2', total_pages)
        
        # [FIX] PMI 전용 프로젝트 정보 수식 연결 및 스타일 설정 (RT 제외)
        if self.tab_name == "PMI":
            try:
                ws0 = wb.worksheets[0]
                if len(wb.worksheets) > 1:
                    # 엑셀 수식을 사용하여 갑지의 값이 바뀌면 자동으로 바뀌게 설정
                    self.safe_set_value(new_sheet, 'K5', f"='{ws0.title}'!L5")
                    self.safe_set_value(new_sheet, 'M5', f"='{ws0.title}'!N5")
                    self.safe_set_value(new_sheet, 'M8', f"='{ws0.title}'!N8")
                    # [FIX] K5:M10 범위 글씨체 바탕, 크기 9 적용
                    for r_idx in range(5, 11):
                        for c_idx in range(11, 14): # K(11) ~ M(13)
                            cell = new_sheet.cell(row=r_idx, column=c_idx)
                            cell.font = Font(name='바탕', size=9, bold=False)
                            if (r_idx == 5 or r_idx == 8) and c_idx == 11: # K5, K8 줄바꿈
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            else:
                                cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
            except: pass
        
        return new_sheet

    def extract_only(self, show_msg=True):
        """데이터만 추출하여 리스트와 미리보기에 반영"""
        target_file = self.target_file_path.get()
        if not target_file:
            messagebox.showwarning("파일 미선택", "데이터 파일을 선택해주세요.")
            return False
            
        self.log(f"🔍 데이터 추출 시작: {os.path.basename(target_file)}")
        self.save_settings()
        
        # [NEW] Extract date from filename
        fname = os.path.basename(target_file)
        # More flexible regex for dates like 2024.03.05, 24.03.05, 24_03_05, 240305, etc.
        date_match = re.search(r'(\d{4}[-._]\d{2}[-._]\d{2}|\d{2}[-._]\d{2}[-._]\d{2}|\d{8}|\d{6})', fname)
        extracted_date = date_match.group(0) if date_match else ""
        
        # Standardize formatting
        if extracted_date:
            # Remove separators like . or _
            clean_date = re.sub(r'[-._]', '', extracted_date)
            if len(clean_date) == 8:
                extracted_date = f"{clean_date[:4]}-{clean_date[4:6]}-{clean_date[6:]}"
            elif len(clean_date) == 6:
                extracted_date = f"20{clean_date[:2]}-{clean_date[2:4]}-{clean_date[4:]}"
        
        if not extracted_date:
            extracted_date = datetime.datetime.now().strftime("%Y-%m-%d")
            self.log(f"⚠️ 파일명에서 날짜를 찾지 못해 오늘 날짜({extracted_date})로 설정합니다.")
        else:
            self.log(f"📅 파일 날짜 인식 성공: {extracted_date}")

        self.progress['value'] = 0
        all_extracted_data = []
        
        try:
            target_input = self.sequence_filter.get().strip()
            target_no_list = [x.strip() for x in target_input.replace(',', ' ').split() if target_input and x.strip()] if target_input else []
            
            with pd.ExcelFile(target_file) as xls:
                for s_idx, sheet_name in enumerate(xls.sheet_names):
                    self.log(f"📄 시트 스캔: {sheet_name}")
                    try: temp_df = pd.read_excel(xls, sheet_name=sheet_name, header=None, nrows=50)
                    except: continue
                    header_idx = None
                    for i, row in temp_df.iterrows():
                        row_str = str(row.values).upper()
                        if self.tab_name == "PMI":
                            if ("CR" in row_str and "NI" in row_str) or ("CHROMIUM" in row_str): header_idx = i; break
                        else:
                            # RT, PT, MT, PAUT 등 일반 NDT 보고서는 주로 NO, ITEM, RESULT, EVALUATION 등의 단어가 포함됨
                            if ("NO" in row_str or "ITEM" in row_str or "SEQ" in row_str) and ("RESULT" in row_str or "ACCEPT" in row_str or "JOINT" in row_str or "SIZE" in row_str or "WELD" in row_str):
                                header_idx = i; break
                    if header_idx is None: continue
                    
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=header_idx)
                    def _find_col(df, keywords):
                        # 1. 우선 정확히 일치하는 컬럼을 찾음
                        for col in df.columns:
                            c_up = str(col).upper().strip()
                            if any(k == c_up for k in keywords): return col
                        # 2. 포함된 컬럼을 찾되, NI의 경우 UNIT 등 오진 가능성 차단
                        for col in df.columns:
                            c_up = str(col).upper().strip()
                            if any(k in c_up for k in keywords):
                                if "NI" in keywords and ("UNIT" in c_up or "LINE" in c_up): continue
                                return col
                        return None
                    col_no = _find_col(df, ["NO.", "NO", "SEQ", "NUM", "POS", "ITEM"])
                    col_joint = _find_col(df, ["JOINT", "J/N", "JOINT NO", "PUNCH", "WELD NO"])
                    col_loc = _find_col(df, ["LOCATION", "TEST POSITION", "POINT", "AREA", "POSITION"])
                    col_dwg = _find_col(df, ["ISO", "DWG", "DRAWING", "LINE"])
                    
                    # NDT 별 고정 열 검출 (PMI)
                    col_cr = col_ni = col_mo = col_mn = col_grade_orig = None
                    if self.tab_name == "PMI":
                        col_cr = _find_col(df, ["CR", "CHROMIUM"]); col_ni = _find_col(df, ["NI", "NICKEL"])
                        col_mo = _find_col(df, ["MO", "MOLYBDENUM"]); col_mn = _find_col(df, ["MN", "MANGANESE"])
                        col_grade_orig = _find_col(df, ["GRADE", "MATERIAL", "SPEC", "TYPE"])

                    # 커스텀(동적) 열 매핑
                    dynamic_col_map = {}
                    for col_idx_cfg, (col_key, col_display, _) in enumerate(self.columns_config):
                        if col_key in ["selected", "No", "Date", "Dwg", "Joint", "Loc", "Ni", "Cr", "Mo", "Mn", "Grade"]: continue
                        
                        # [FIX] Welder No 등 유사 키워드 대응 강화
                        keywords = [col_key.upper(), col_display.upper().replace(" ", "")]
                        if col_key == "Welder": keywords.extend(["WELDER NO", "WELD NO", "WELD"])
                        if col_key == "Accept": keywords.extend(["ACC"])
                        if col_key == "Remarks": keywords.extend(["REMARK", "비고", "NOTE"])
                        
                        found_c = _find_col(df, keywords)
                        if found_c is not None:
                            dynamic_col_map[col_key] = found_c

                    for _, row in df.iterrows():
                        v_raw_no = str(row[col_no]).strip() if col_no is not None else str(_+1)
                        if v_raw_no.lower() == "nan": v_raw_no = ""
                        if target_no_list and v_raw_no not in target_no_list: continue
                        
                        row_data = {
                            'No': v_raw_no,
                            'Joint': str(row[col_joint]).strip() if col_joint is not None else v_raw_no,
                            'Loc': str(row[col_loc]).strip() if col_loc is not None else "",
                            'Dwg': str(row[col_dwg]).strip() if col_dwg is not None else "",
                            'Date': extracted_date,
                            'selected': True
                        }
                        
                        # 'nan' 무시 처리
                        for k in ['Joint', 'Loc', 'Dwg']:
                            if row_data[k].lower() == "nan": row_data[k] = ""
                        
                        if self.tab_name == "PMI":
                            v_cr = self.to_float(row[col_cr]) if col_cr is not None else 0.0
                            if v_cr > 0 or (v_raw_no != ""):
                                v_ni = self.to_float(row[col_ni]) if col_ni is not None else 0.0
                                v_mo = self.to_float(row[col_mo]) if col_mo is not None else 0.0
                                v_mn = self.to_float(row[col_mn]) if col_mn is not None else 0.0
                                orig_grade = str(row[col_grade_orig]).strip() if col_grade_orig is not None else ""
                                if orig_grade.lower() == "nan": orig_grade = ""
                                
                                final_grade = orig_grade
                                if self.auto_verify and self.auto_verify.get():
                                    detected = self.check_material_grade({'Cr': v_cr, 'Ni': v_ni, 'Mo': v_mo, 'Mn': v_mn})
                                    if detected: final_grade = detected
                                if not final_grade or final_grade == "nan":
                                    final_grade = "SS316" if v_mo >= 1.5 else "SS304"
                                    
                                row_data.update({'Cr': v_cr, 'Ni': v_ni, 'Mo': v_mo, 'Mn': v_mn, 'Grade': final_grade})
                                
                                # 동적 컬럼 채우기
                                for col_key, df_col in dynamic_col_map.items():
                                    val = str(row[df_col]).strip()
                                    row_data[col_key] = "" if val.lower() == "nan" else val
                                    
                                all_extracted_data.append(row_data)
                        else:
                            # RT, PT, MT, PAUT 로직 (PMI 이외)
                            if v_raw_no != "":
                                # 동적 컬럼 채우기
                                for col_key, df_col in dynamic_col_map.items():
                                    val = str(row[df_col]).strip()
                                    row_data[col_key] = "" if val.lower() == "nan" else val
                                
                                # Result 컬럼의 기본값 통과(ACCEPT) 처리 (데이터에 없고 컬럼 스키마에만 있는 경우)
                                if 'Result' in self.column_keys and not row_data.get('Result'):
                                    row_data['Result'] = "ACCEPT"
                                    
                                all_extracted_data.append(row_data)

                    # [NEW] 원소 함량 필터 적용
                    if self.tab_name == "PMI" and self.element_filters:
                        filtered_data = []
                        for item in all_extracted_data:
                            keep = True
                            for f_opt in self.element_filters:
                                f_key = f_opt['key'].get().strip()
                                f_min = f_opt['min'].get().strip()
                                f_max = f_opt['max'].get().strip()
                                
                                if not f_key: continue
                                
                                val = self.to_float(item.get(f_key, 0.0))
                                
                                if f_min:
                                    try:
                                        if val < float(f_min): keep = False; break
                                    except ValueError: pass
                                if f_max:
                                    try:
                                        if val > float(f_max): keep = False; break
                                    except ValueError: pass
                            
                            if keep:
                                filtered_data.append(item)
                        
                        if len(filtered_data) < len(all_extracted_data):
                            self.log(f"✂️ 함량 필터 적용: {len(all_extracted_data)}개 -> {len(filtered_data)}개 추출")
                        all_extracted_data = filtered_data
                    self.progress['value'] = ((s_idx + 1) / len(xls.sheet_names)) * 50

            if not all_extracted_data:
                messagebox.showerror("오류", "추출된 데이터가 없습니다.")
                return False

            # Extraction Mode Filter
            mode = self.extraction_mode.get()
            if mode != "전체":
                original_count = len(all_extracted_data)
                if mode == "SS304 만": all_extracted_data = [d for d in all_extracted_data if d['Grade'] == "SS304"]
                elif mode == "SS316 만": all_extracted_data = [d for d in all_extracted_data if d['Grade'] == "SS316"]
                elif mode == "DUPLEX 만": all_extracted_data = [d for d in all_extracted_data if d['Grade'] == "DUPLEX"]
                elif mode == "SS310 만": all_extracted_data = [d for d in all_extracted_data if d['Grade'] == "SS310"]
                elif mode == "미분류(기타) 만":
                    known_grades = ["SS304", "SS316", "DUPLEX", "SS310"]
                    all_extracted_data = [d for d in all_extracted_data if d['Grade'] not in known_grades]
                
                self.log(f"🔍 필터링 ({mode}): {original_count}개 -> {len(all_extracted_data)}개")
                if not all_extracted_data:
                    messagebox.showinfo("알림", f"'{mode}'에 해당하는 데이터가 없습니다.")
                    return False

            if target_no_list:
                all_extracted_data.sort(key=lambda x: target_no_list.index(str(x['No'])) if str(x['No']) in target_no_list else 999999)

            # [CHANGE] Overwrite -> Accumulate
            self.extracted_data.extend(all_extracted_data)
            self.populate_preview(self.extracted_data)
            self.update_date_listbox()
            self.progress['value'] = 100
            if show_msg:
                self.log(f"✅ 데이터 누적 완료 (현재 총 {len(self.extracted_data)} 건)")
                messagebox.showinfo("완료", f"데이터가 추가되었습니다.\n현재 목록에는 총 {len(self.extracted_data)}건의 데이터가 있습니다.")
            return True
        except Exception as e:
            self.log(f"❌ 추출 오류: {e}")
            traceback.print_exc()
            return False

    def run_process(self):
        target_file = self.target_file_path.get()
        template_path = self.template_file_path.get()
        
        # [MOD] Only require template if extracted_data is already present. 
        # If extracted_data is empty, target_file is needed to perform extraction.
        if not template_path:
            messagebox.showwarning("파일 미선택", "성적서 양식 파일을 선택해주세요.")
            return

        if not self.extracted_data and not target_file:
            messagebox.showwarning("파일 미선택", "데이터 추출을 위해 원본 데이터 파일을 선택해주세요.")
            return
            
        self.save_settings()
        
        # 데이터가 비어있는 경우에만 추출 수행
        if not self.extracted_data:
            if not self.extract_only(show_msg=False): return
            
        # [NEW] 체크된 항목만 필터링 (기본값은 True)
        final_list = [d for d in self.extracted_data if d.get('selected', True)]
        if not final_list:
            messagebox.showwarning("항목 미선택", "선택된 데이터가 없습니다. 미리보기에서 항목을 체크해주세요.")
            return

        self.log(f"🚀 성적서 생성 시작 (총 {len(final_list)} 건)...")
        self.progress['value'] = 0
        
        try:
            all_extracted_data = final_list
            # [STRICT] RT의 경우 항상 11행부터 시작하도록 강제함 (저장된 설정 무시 목적)
            if self.tab_name == "RT":
                self.config['START_ROW'] = 11
            
            rows_per_item = 3 if self.tab_name == "PMI" else 1
            self.log(f"📊 설정 확인: {self.tab_name} 시작행={self.config['START_ROW']}, 개당 행수={rows_per_item}")
            total_items = len(all_extracted_data)
            max_rows_per_page = self.config['DATA_END_ROW'] - self.config['START_ROW'] + 1
            items_per_page = max_rows_per_page // rows_per_item
            total_pages = math.ceil(total_items / items_per_page)
            if total_pages < 1: total_pages = 1
            total_report_pages = total_pages + 1 # 갑지 포함

            wb = openpyxl.load_workbook(template_path, keep_vba=True)
            if len(wb.worksheets) >= 1:
                ws0 = wb.worksheets[0]
                # [NEW] 갑지 여백 조절을 위한 행 높이 설정
                ws0.row_dimensions[1].height = 25 # 상단 여백
                bottom_row = 40 if self.tab_name == "RT" else self.config.get('PRINT_END_ROW', 45)
                ws0.row_dimensions[bottom_row].height = 25 # 하단 여백
                
                self.add_logos_to_sheet(ws0, is_cover=True, clear_existing=False)
                
                if self.tab_name == "RT":
                    # [MOD] RT 성적서 갑지 페이지 번호 형식 (Page 1 of N)
                    # N3와 P3가 병합된 경우를 고려하여 N3에만 값을 넣고 P3는 건드리지 않음. 가운데 정렬 적용.
                    self.safe_set_value(ws0, 'N3', f"Page   1  of   {total_report_pages}", align='center')
                else:
                    self.safe_set_value(ws0, 'N3', 1) 
                    self.safe_set_value(ws0, 'P3', total_report_pages)
                # 불필요하거나 이미지와 간섭을 줄 수 있는 보더 초기화 제거 (필요 시에만 사용)
                # for r in range(23, 39): ...
            
            data_sheet_id = 1 if len(wb.worksheets) >= 2 else 0
            ws = wb.worksheets[data_sheet_id]; ws.title = f"{ws.title[:20]}_001"
            # [RT/NDT] 템플릿의 그림/바닥글 보호를 위해 clear_existing=False
            self.add_logos_to_sheet(ws, is_cover=False, clear_existing=(self.tab_name == "PMI"))
            self.force_print_settings(ws); self.set_eulji_headers(ws)
            # 을지 첫 장 페이지 번호
            if self.tab_name == "RT":
                # [MOD] RT 성적서 을지 첫 장 페이지 번호 (표지가 1이므로 을지는 2부터 시작)
                self.safe_set_value(ws, 'R2', f"Page   2  of   {total_report_pages}")
            else:
                self.safe_set_value(ws, 'R2', 2)
                self.safe_set_value(ws, 'T2', total_report_pages)
            
            if self.tab_name == "PMI":
                try:
                    if len(wb.worksheets) >= 2:
                        # [FIX] 수식으로 연결 (갑지 L5 -> 을지 K5, 갑지 N5 -> 을지 M5, 갑지 N8 -> 을지 M8)
                        self.safe_set_value(ws, 'K5', f"='{ws0.title}'!L5")
                        self.safe_set_value(ws, 'M5', f"='{ws0.title}'!N5")
                        self.safe_set_value(ws, 'M8', f"='{ws0.title}'!N8")
                        # [FIX] K5:M10 범위 스타일 설정 (바탕, 크기 9)
                        for r_idx in range(5, 11):
                            for c_idx in range(11, 14):
                                cell = ws.cell(row=r_idx, column=c_idx)
                                cell.font = Font(name='바탕', size=9, bold=False)
                                if (r_idx == 5 or r_idx == 8) and c_idx == 11: # K5, K8 줄바꿈
                                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                else:
                                    cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                except Exception as e:
                    self.log(f"갑지 데이터 복사 실패: {e}")
            
            materials = "SS304,SS304L,SS316,SS316L,SS321,SS347,SS410,SS430,DUPLEX,MONEL,INCONEL,ER308,ER308L,ER309,ER309L,ER316,ER316L,ER347,ER2209,WP316,WP316L,TP316,TP316L,F316L,A182-F316L,A312-TP316L"
            dv_q = DataValidation(type="list", formula1=f'"{materials}"', allow_blank=True)
            # [NEW] 데이터 기입 전 기존 병합 영역 정리 (A-E 등 중복 병합 방지)
            def clear_merges_in_range(sheet, start_row, end_row):
                merged_ranges = list(sheet.merged_cells.ranges)
                for r in merged_ranges:
                    # 영역이 겹치면 해제 (aggressive clearing)
                    if r.min_row <= end_row and r.max_row >= start_row:
                        try: sheet.unmerge_cells(str(r))
                        except: pass

            if self.tab_name == "PMI":
                ws = wb.worksheets[data_sheet_id]
                clear_merges_in_range(ws, self.config['START_ROW'], self.config['DATA_END_ROW'] + 20)
                ws.add_data_validation(dv_q)
            else:
                ws = wb.worksheets[data_sheet_id]
                # RT의 경우 템플릿의 병합 상태를 유지하기 위해 unmerge 생략
                
            # [NEW] 템플릿의 헤더를 읽어 동적 컬럼 매핑 생성
            excel_col_map = {}
            if self.tab_name != "PMI":
                # [MOD] RT의 경우 9~10행 모두 스캔하여 원문자 기호 매핑 정확도 향상
                header_rows = [9, 10] if self.tab_name == "RT" else ([self.config['START_ROW'] - 1] if self.config['START_ROW'] > 1 else [])
                
                for header_row in header_rows:
                    if header_row <= 0: continue
                    # [MOD] A(1) 부터 X(24) 열까지 전체 스캔 (A, B열 유동성 확보)
                    for c_idx in range(1, 26): 
                        cell = ws.cell(row=header_row, column=c_idx)
                        cell_val = cell.value
                        
                        # [FIX] 병합된 셀인 경우 기준 셀의 값 가져오기
                        if cell_val is None:
                            for m_range in ws.merged_cells.ranges:
                                if cell.coordinate in m_range:
                                    cell_val = ws.cell(row=m_range.min_row, column=m_range.min_col).value
                                    break
                                    
                        if not cell_val: continue
                        # 표준화 (공백, 줄바꿈, 마침표 제거)
                        header_str = str(cell_val).upper().replace("\n", "").replace(" ", "").replace(".", "")
                        
                        # 미리보기 컬럼들 중 헤더와 일치하는 것이 있는지 확인
                        for col_key, col_display, _ in self.columns_config:
                            # A, B 고정 매핑항목(Joint, Loc) 및 No, Date 등은 제외
                            if col_key in ["selected", "No", "Date", "Dwg", "Joint", "Loc"]: continue
                            
                            # 내부 키와 표시 이름을 모두 비교 (표준화 후)
                            k_up = str(col_key).upper().replace(" ", "").replace(".", "")
                            d_up = str(col_display).upper().replace(" ", "").replace(".", "")
                            
                            # [NEW] Remarks 컬럼 매핑 보강 (비고, REMARK 등)
                            if col_key == "Remarks":
                                if any(x in header_str for x in ["REMARK", "비고", "NOTE"]):
                                    if col_key not in excel_col_map:
                                        excel_col_map[col_key] = c_idx
                                        continue

                            if k_up in header_str or header_str in k_up or d_up in header_str or header_str in d_up:
                                # 이미 선점된 컬럼이 아니면 매핑
                                if col_key not in excel_col_map:
                                    excel_col_map[col_key] = c_idx
                            
                            # [NEW] 원문자(circled numbers) 매핑 보강 (RT 전용)
                            elif self.tab_name == "RT" and col_key.startswith("D") and col_key[1:].isdigit():
                                d_idx = int(col_key[1:])
                                circled_nums = ["①","②","③","④","⑤","⑥","⑦","⑧","⑨","⑩","⑪","⑫","⑬","⑭","⑮"]
                                if 1 <= d_idx <= 15:
                                    symbol = circled_nums[d_idx-1]
                                    if symbol in header_str:
                                        if col_key not in excel_col_map:
                                            excel_col_map[col_key] = c_idx
                if excel_col_map:
                    self.log(f"🔗 자동 매핑 완료: {', '.join([f'{k}({v})' for k,v in excel_col_map.items()])}")
                else:
                    self.log("⚠️ 템플릿에서 동적 컬럼(Welder No 등)을 찾지 못했습니다. 헤더 행(9-10행)을 확인해주세요.")

            rows_per_item = 3 if self.tab_name == "PMI" else 1
            current_row = self.config['START_ROW']; current_page = 1; data_ptr = 0
            
            while data_ptr < len(all_extracted_data):
                # 블록 크기만큼 남은 공간이 있는지 확인
                if current_row + rows_per_item - 1 > self.config['DATA_END_ROW']:
                    current_page += 1; ws = self.prepare_next_sheet(wb, data_sheet_id, current_page + 1, total_report_pages)
                    if self.tab_name == "PMI":
                        clear_merges_in_range(ws, self.config['START_ROW'], self.config['DATA_END_ROW'] + 20)
                        ws.add_data_validation(dv_q)
                    current_row = self.config['START_ROW']

                # 현재 블록/행에 들어갈 데이터 (PMI는 3개, RT는 1헤더+3개(최대), 나머지는 1개)
                batch_size = 3 if self.tab_name == "PMI" else 1
                batch = all_extracted_data[data_ptr : data_ptr + batch_size]
                
                # 테두리 및 서식 적용
                for r_offset in range(rows_per_item):
                    r = current_row + r_offset
                    if self.tab_name == "PMI":
                        rd = ws.row_dimensions[r]; rd.height = 20.55
                    last_col = 13 # RT 테두리 요청 취소: 기존처럼 13열(M)까지만 적용하거나, 템플릿 선 유지
                    for c in range(1, last_col + 1):
                        cell = ws.cell(row=r, column=c)
                        l_s = thin_side; r_s = thin_side
                        t_s = medium_side if r == self.config['START_ROW'] else thin_side
                        b_s = medium_side if r == self.config['DATA_END_ROW'] else thin_side
                        
                        if c == 1: l_s = medium_side
                        if c == last_col: r_s = medium_side
                        
                        # [PMI 전용] A-E 수직 병합 구역 내부 선 제거
                        if rows_per_item > 1 and c <= 5: 
                            if 1 < c < 5: r_s = Side(style=None); l_s = Side(style=None)
                            elif c == 1: r_s = Side(style=None)
                            elif c == 5: l_s = Side(style=None)
                            if r_offset < rows_per_item - 1: b_s = Side(style=None)
                            if r_offset > 0: t_s = Side(style=None)
                        # [MOD] RT 성적서의 경우 프로그램에서 테두리를 그리지 않고 템플릿의 선을 그대로 사용함
                        if self.tab_name != "RT":
                            cell.border = Border(left=l_s, right=r_s, top=t_s, bottom=b_s)

                if rows_per_item > 1:
                    if self.tab_name == "PMI":
                        # [PMI] A-E 수직 병합 (Drawing No)
                        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + rows_per_item - 1, end_column=5)
                        cell_a = ws.cell(row=current_row, column=1); self.safe_set_value(ws, cell_a.coordinate, batch[0].get('Dwg', ''))
                        cell_a.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                        # [PMI] F열 수직 병합 (Joint No)
                        ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row + rows_per_item - 1, end_column=6)
                        cell_f_merged = ws.cell(row=current_row, column=6); self.safe_set_value(ws, cell_f_merged.coordinate, batch[0].get('Joint', batch[0].get('No', '')))
                        cell_f_merged.alignment = Alignment(horizontal='center', vertical='center')
                    elif self.tab_name == "RT":
                        # [RT] 4행 1블록 (1행은 Drawing No 헤더)
                        cell_h = ws.cell(row=current_row, column=1)
                        # 원본 서식(Drawing No. ) 유지 시도
                        orig_val = str(cell_h.value or "")
                        dwg_val = batch[0].get('Dwg', '')
                        if "Drawing" in orig_val:
                            # 접두어 유지 (이미 합쳐진 경우 중복 방지)
                            prefix = orig_val.split(":")[0] + ":" if ":" in orig_val else (orig_val.split(".")[0] + "." if "." in orig_val else "Drawing No.")
                            new_val = f"{prefix} {dwg_val}"
                        else:
                            new_val = f"Drawing No. {dwg_val}"
                        self.safe_set_value(ws, cell_h.coordinate, new_val)
                        # [FIX] 왼쪽 정렬을 적용하고 B열을 비워야 옆 칸으로 글자가 넘어감 (Overflow)
                        cell_h.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    # [RT/NDT] 1행 1데이터
                    # A열: Joint
                    cell_a = ws.cell(row=current_row, column=1); self.safe_set_value(ws, cell_a.coordinate, batch[0].get('Joint', batch[0].get('No', '')))
                    # B열: Loc
                    cell_b = ws.cell(row=current_row, column=2); self.safe_set_value(ws, cell_b.coordinate, batch[0].get('Loc', ''))
                    
                    # C열 이후는 excel_col_map에 따라 동적 기입 (Accept 등 포함)
                    
                    # [MOD] RT에서는 A-E 병합을 하지 않음 (각 컬럼별 데이터 기입을 위해)
                    # 기존 A-E 병합 코드는 주석 처리하거나 조건부로 실행하지 않음

                # 개별 데이터 입력
                for i, item in enumerate(batch):
                    # RT는 첫 행이 헤더이므로 데이터는 한 행씩 밀려서 작성 (PMI는 그대로)
                    r = current_row + i
                    
                    if self.tab_name == "PMI":
                        cell_g = ws.cell(row=r, column=7); self.safe_set_value(ws, cell_g.coordinate, item.get('Loc', ''))
                        cell_g.alignment = Alignment(horizontal='center', vertical='center')
                        for val_key, col_idx in [('Ni', 8), ('Cr', 9), ('Mo', 10)]:
                            v = item.get(val_key, 0.0); cell = ws.cell(row=r, column=col_idx)
                            self.safe_set_value(ws, cell.coordinate, v if v > 0 else "")
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            if v > 0: cell.number_format = '0.00'
                        cell_l = ws.cell(row=r, column=13); self.safe_set_value(ws, cell_l.coordinate, item.get('Grade', ''))
                        cell_l.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); cell_l.font = Font(size=8.5); dv_q.add(cell_l)
                    else:
                        # [RT] 고정 컬럼 (A, B) 및 오버플로우 지원
                        if self.tab_name != "RT":
                            # PMI 등 기존 탭은 A, B열 고정 기입 유지
                            cell_joint = ws.cell(row=r, column=1); self.safe_set_value(ws, cell_joint.coordinate, item.get('Joint', item.get('No', '')))
                            cell_loc = ws.cell(row=r, column=2); self.safe_set_value(ws, cell_loc.coordinate, item.get('Loc', ''))
                            
                            if self.tab_name == "PMI":
                                cell_joint.alignment = Alignment(horizontal='center', vertical='center')
                                cell_loc.alignment = Alignment(horizontal='center', vertical='center')
                        
                        for col_key in self.column_keys:
                            # [FIX] Accept/Reject 컬럼 명시적 기입 허용 (RT 전용)
                            if col_key in ["selected", "No", "Date"]: continue
                            # PMI는 기존처럼 별도 블록에서 처리하므로 제외
                            if self.tab_name == "PMI" and col_key in ["Dwg", "Joint", "Loc"]: continue
                            
                            c_idx = excel_col_map.get(col_key)
                            if c_idx:
                                cell = ws.cell(row=r, column=c_idx); self.safe_set_value(ws, cell.coordinate, item.get(col_key, ''))
                                # [MOD] RT 역시 기본은 가운데 정렬, Drawing No만 왼쪽 정렬(오버플로우 및 잘림 방지)
                                if self.tab_name == "RT":
                                    align = 'left' if col_key == 'Dwg' else 'center'
                                    cell.alignment = Alignment(horizontal=align, vertical='center')
                                elif self.tab_name == "PMI":
                                    cell.alignment = Alignment(horizontal='center', vertical='center')

                data_ptr += len(batch); current_row += rows_per_item; self.progress['value'] = 30 + (data_ptr / len(all_extracted_data)) * 65

            # 빈 칸 채우기 (데이터 끝난 후부터 DATA_END_ROW까지)
            while current_row + rows_per_item - 1 <= self.config['DATA_END_ROW']:
                for r_offset in range(rows_per_item):
                    r = current_row + r_offset
                    if self.tab_name == "PMI":
                        rd = ws.row_dimensions[r]; rd.height = 20.55
                    for c in range(1, 14): # A(1)~M(13) 고정
                        cell = ws.cell(row=r, column=c)
                        l_s = thin_side; r_s = thin_side
                        t_s = medium_side if r == self.config['START_ROW'] else thin_side
                        b_s = medium_side if r == self.config['DATA_END_ROW'] else thin_side
                        if c == 1: l_s = medium_side
                        if c == 13: r_s = medium_side
                        if rows_per_item > 1 and c <= 5: # PMI 전용 상하 선 제거
                            if 1 < c < 5: r_s = Side(style=None); l_s = Side(style=None)
                            elif c == 1: r_s = Side(style=None)
                            elif c == 5: l_s = Side(style=None)
                            if r_offset < rows_per_item - 1: b_s = Side(style=None)
                            if r_offset > 0: t_s = Side(style=None)
                        # [MOD] RT 성적서 테두리 그리기 생략
                        if self.tab_name != "RT":
                            cell.border = Border(left=l_s, right=r_s, top=t_s, bottom=b_s)
                
                # 빈 블록 병합 (PMI 전용 혹은 NDTTab.rows_per_item > 1 인 경우)
                if rows_per_item > 1:
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + rows_per_item - 1, end_column=5)
                if rows_per_item > 1:
                    ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row + rows_per_item - 1, end_column=6)
                
                if self.tab_name == "PMI": dv_q.add(ws.cell(row=current_row, column=13))
                current_row += rows_per_item

            # [FORCE] 모든 데이터 시트의 45행 바닥선(A-L) 설정 (PMI 전용)
            if self.tab_name != "RT":
                data_end_row = int(self.config.get('DATA_END_ROW', 45))
                for idx, s in enumerate(wb.worksheets):
                    if s.max_row >= data_end_row:
                        for c in range(1, 14):
                            cell = s.cell(row=data_end_row, column=c)
                            curr_border = cell.border
                            l_s = curr_border.left; r_s = curr_border.right; t_s = curr_border.top
                            
                            if idx == 0: # 갑지
                                if c in [1, 2, 3, 11, 13]: # A, B, C, K, M 바닥선 제거
                                    b_s = Side(style=None)
                                else: # D~J 는 약한선
                                    b_s = thin_side
                            else: # 을지
                                b_s = medium_side
                                
                            cell.border = Border(left=l_s, right=r_s, top=t_s, bottom=b_s)

            # [FORCE] 을지 시트의 지정 행(17~45) 글꼴 크기 통일 (PMI 전용)
            if self.tab_name != "RT":
                for idx, s in enumerate(wb.worksheets):
                    if idx > 0: # 을지 시트
                        for r_idx in range(17, data_end_row + 1):
                            for c_idx in range(1, 14):
                                cell = s.cell(row=r_idx, column=c_idx)
                                f = cell.font
                                if f:
                                    cell.font = Font(name=f.name, size=10, bold=f.bold, italic=f.italic, vertAlign=f.vertAlign, underline=f.underline, strike=f.strike, color=f.color)
                                else:
                                    cell.font = Font(name='맑은 고딕', size=10)

            # [REMOVED] 중복되고 오류 가능성 있는 하단 페이지 번호 루프 제거 (이미 prepare_next_sheet 및 run_process 상단에서 처리됨)

            now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_name = f"{os.path.splitext(os.path.basename(template_path))[0]}_Unified_{now_str}{os.path.splitext(template_path)[1]}"
            save_path = os.path.join(os.path.dirname(template_path), output_name); wb.save(save_path)
            wb.close() # [FIX] Release file handles, especially for keep_vba=True
            self.progress['value'] = 100; self.log(f"✨ 완료! 저장됨: {output_name}")
            messagebox.showinfo("성공", f"통합 성적서 생성이 완료되었습니다.\n\n경로: {os.path.dirname(save_path)}\n파일명: {output_name}")
        except Exception as e: self.log(f"❌ 오류: {e}"); traceback.print_exc()
        finally:
            for f in glob.glob(os.path.join(tempfile.gettempdir(), "temp_*.png")):
                try: os.remove(f)
                except: pass

class MainApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PMI 및 NDT 성적서 자동 생성 통합 도구")
        self.root.geometry("850x800")
        self.root.configure(background="#f9fafb")
        
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TFrame", background="#f9fafb")
        style.configure("TLabel", background="#f9fafb", font=("Malgun Gothic", 10))
        style.configure("TLabelframe", background="#f9fafb", font=("Malgun Gothic", 10, "bold"))
        style.configure("TLabelframe.Label", background="#f9fafb", font=("Malgun Gothic", 10, "bold"))
        style.configure("Action.TButton", font=("Malgun Gothic", 11, "bold"), padding=10)
        style.configure("Main.TNotebook.Tab", font=("Malgun Gothic", 11, "bold"), padding=[10, 5])
        
        self.main_notebook = ttk.Notebook(self.root, style="Main.TNotebook")
        self.main_notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.tabs = {}
        # Create tabs using the same NDTTab class for all modes
        for mode in ["PMI", "RT", "PT", "MT", "PAUT"]:
            self.tabs[mode] = NDTTab(self.main_notebook, mode, self.root)

if __name__ == "__main__":
    root = tk.Tk()
    app = MainApp(root)
    root.mainloop()
