
import sys
import os
import glob
import math
import traceback
import re
import warnings
import json
import tempfile
from tkinter import filedialog, Tk, Label, Entry, Button, Frame, StringVar, ttk
from tkinter import messagebox, simpledialog

# 필수 라이브러리 직접 import (PyInstaller 호환)
import pandas as pd
import openpyxl 
import datetime 
from openpyxl.cell.cell import MergedCell 
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.image import Image as XLImage 
from openpyxl.worksheet.datavalidation import DataValidation 
from openpyxl.styles import Alignment, Font  # ★ 줄바꿈 및 글자 크기 기능을 위해 필수
from PIL import Image as PILImage, ImageChops 

from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# 경고 메시지 무시
warnings.simplefilter("ignore")

# ========================================================
# [설정] 전역 변수 초기화
# ========================================================

# [1] 갑지 (Cover) 변수 - logo_settings.json 기본값 내장
SEOUL_COVER_ANCHOR="E5"; SEOUL_COVER_W=200.0; SEOUL_COVER_H=18.0; SEOUL_COVER_X=30.0; SEOUL_COVER_Y=15.0
SITCO_COVER_ANCHOR="A6"; SITCO_COVER_W=80.0; SITCO_COVER_H=40.0; SITCO_COVER_X=15.0; SITCO_COVER_Y=10.0
FOOTER_COVER_ANCHOR="Q51"; FOOTER_COVER_W=80.0; FOOTER_COVER_H=20.0; FOOTER_COVER_X=-10.0; FOOTER_COVER_Y=5.0
FOOTER_PT_COVER_ANCHOR="A51"; FOOTER_PT_COVER_W=100.0; FOOTER_PT_COVER_H=25.0; FOOTER_PT_COVER_X=3.0; FOOTER_PT_COVER_Y=5.0

# [2] 을지 (Data) 변수 - logo_settings.json 기본값 내장
SEOUL_DATA_ANCHOR="F5"; SEOUL_DATA_W=200.0; SEOUL_DATA_H=18.0; SEOUL_DATA_X=35.0; SEOUL_DATA_Y=15.0
SITCO_DATA_ANCHOR="A6"; SITCO_DATA_W=80.0; SITCO_DATA_H=40.0; SITCO_DATA_X=5.0; SITCO_DATA_Y=10.0
FOOTER_DATA_ANCHOR="Q37"; FOOTER_DATA_W=100.0; FOOTER_DATA_H=15.0; FOOTER_DATA_X=5.0; FOOTER_DATA_Y=3.0
FOOTER_PT_DATA_ANCHOR="A37"; FOOTER_PT_DATA_W=100.0; FOOTER_PT_DATA_H=30.0; FOOTER_PT_DATA_X=3.0; FOOTER_PT_DATA_Y=-10.0

# [3] 행 설정
START_ROW = 20       
DATA_END_ROW = 35    
PRINT_END_ROW = 37   

PAGE_NUMBER_CELL = "Q3"  # 을지 시트 페이지 번호 셀

# 경로 설정 - 스크립트 원본 위치를 고정 (다른 곳에서 실행해도 설정 유지)
if getattr(sys, 'frozen', False):
    # PyInstaller EXE 실행 시: EXE 파일 위치 기준
    SCRIPT_HOME = os.path.dirname(sys.executable)
elif '__file__' in globals():
    # Python 소스 실행 시: .py 파일 위치 기준
    SCRIPT_HOME = os.path.dirname(os.path.abspath(__file__))
else:
    # 대화형 모드: 현재 작업 디렉터리
    SCRIPT_HOME = os.getcwd()

folder_path = SCRIPT_HOME  # 기본 폴더 경로
logo_folder_path = SCRIPT_HOME  # 로고 파일 검색 경로 (초기값)
SETTINGS_FILE = os.path.join(SCRIPT_HOME, "logo_settings.json")  # 설정 파일은 항상 스크립트 원본 위치에 저장

# ========================================================
# [설정 저장/불러오기]
# ========================================================
def load_settings():
    global SEOUL_COVER_X, SEOUL_COVER_Y, SEOUL_COVER_W, SEOUL_COVER_H, SEOUL_COVER_ANCHOR
    global SITCO_COVER_X, SITCO_COVER_Y, SITCO_COVER_W, SITCO_COVER_H, SITCO_COVER_ANCHOR
    global FOOTER_COVER_X, FOOTER_COVER_Y, FOOTER_COVER_W, FOOTER_COVER_H, FOOTER_COVER_ANCHOR
    global FOOTER_PT_COVER_X, FOOTER_PT_COVER_Y, FOOTER_PT_COVER_W, FOOTER_PT_COVER_H, FOOTER_PT_COVER_ANCHOR
    
    global SEOUL_DATA_X, SEOUL_DATA_Y, SEOUL_DATA_W, SEOUL_DATA_H, SEOUL_DATA_ANCHOR
    global SITCO_DATA_X, SITCO_DATA_Y, SITCO_DATA_W, SITCO_DATA_H, SITCO_DATA_ANCHOR
    global FOOTER_DATA_X, FOOTER_DATA_Y, FOOTER_DATA_W, FOOTER_DATA_H, FOOTER_DATA_ANCHOR
    global FOOTER_PT_DATA_X, FOOTER_PT_DATA_Y, FOOTER_PT_DATA_W, FOOTER_PT_DATA_H, FOOTER_PT_DATA_ANCHOR
    
    global START_ROW, DATA_END_ROW, PRINT_END_ROW

    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 갑지
            SEOUL_COVER_ANCHOR = data.get('SEOUL_COVER_ANCHOR', SEOUL_COVER_ANCHOR)
            SEOUL_COVER_X = data.get('SEOUL_COVER_X', SEOUL_COVER_X)
            SEOUL_COVER_Y = data.get('SEOUL_COVER_Y', SEOUL_COVER_Y)
            SEOUL_COVER_W = data.get('SEOUL_COVER_W', SEOUL_COVER_W)
            SEOUL_COVER_H = data.get('SEOUL_COVER_H', SEOUL_COVER_H)

            SITCO_COVER_ANCHOR = data.get('SITCO_COVER_ANCHOR', SITCO_COVER_ANCHOR)
            SITCO_COVER_X = data.get('SITCO_COVER_X', SITCO_COVER_X)
            SITCO_COVER_Y = data.get('SITCO_COVER_Y', SITCO_COVER_Y)
            SITCO_COVER_W = data.get('SITCO_COVER_W', SITCO_COVER_W)
            SITCO_COVER_H = data.get('SITCO_COVER_H', SITCO_COVER_H)

            FOOTER_COVER_ANCHOR = data.get('FOOTER_COVER_ANCHOR', FOOTER_COVER_ANCHOR)
            FOOTER_COVER_X = data.get('FOOTER_COVER_X', FOOTER_COVER_X)
            FOOTER_COVER_Y = data.get('FOOTER_COVER_Y', FOOTER_COVER_Y)
            FOOTER_COVER_W = data.get('FOOTER_COVER_W', FOOTER_COVER_W)
            FOOTER_COVER_H = data.get('FOOTER_COVER_H', FOOTER_COVER_H)

            FOOTER_PT_COVER_ANCHOR = data.get('FOOTER_PT_COVER_ANCHOR', FOOTER_PT_COVER_ANCHOR)
            FOOTER_PT_COVER_X = data.get('FOOTER_PT_COVER_X', FOOTER_PT_COVER_X)
            FOOTER_PT_COVER_Y = data.get('FOOTER_PT_COVER_Y', FOOTER_PT_COVER_Y)
            FOOTER_PT_COVER_W = data.get('FOOTER_PT_COVER_W', FOOTER_PT_COVER_W)
            FOOTER_PT_COVER_H = data.get('FOOTER_PT_COVER_H', FOOTER_PT_COVER_H)

            # 을지
            SEOUL_DATA_ANCHOR = data.get('SEOUL_DATA_ANCHOR', SEOUL_DATA_ANCHOR)
            SEOUL_DATA_X = data.get('SEOUL_DATA_X', SEOUL_DATA_X)
            SEOUL_DATA_Y = data.get('SEOUL_DATA_Y', SEOUL_DATA_Y)
            SEOUL_DATA_W = data.get('SEOUL_DATA_W', SEOUL_DATA_W)
            SEOUL_DATA_H = data.get('SEOUL_DATA_H', SEOUL_DATA_H)

            SITCO_DATA_ANCHOR = data.get('SITCO_DATA_ANCHOR', SITCO_DATA_ANCHOR)
            SITCO_DATA_X = data.get('SITCO_DATA_X', SITCO_DATA_X)
            SITCO_DATA_Y = data.get('SITCO_DATA_Y', SITCO_DATA_Y)
            SITCO_DATA_W = data.get('SITCO_DATA_W', SITCO_DATA_W)
            SITCO_DATA_H = data.get('SITCO_DATA_H', SITCO_DATA_H)

            FOOTER_DATA_ANCHOR = data.get('FOOTER_DATA_ANCHOR', FOOTER_DATA_ANCHOR)
            FOOTER_DATA_X = data.get('FOOTER_DATA_X', FOOTER_DATA_X)
            FOOTER_DATA_Y = data.get('FOOTER_DATA_Y', FOOTER_DATA_Y)
            FOOTER_DATA_W = data.get('FOOTER_DATA_W', FOOTER_DATA_W)
            FOOTER_DATA_H = data.get('FOOTER_DATA_H', FOOTER_DATA_H)

            FOOTER_PT_DATA_ANCHOR = data.get('FOOTER_PT_DATA_ANCHOR', FOOTER_PT_DATA_ANCHOR)
            FOOTER_PT_DATA_X = data.get('FOOTER_PT_DATA_X', FOOTER_PT_DATA_X)
            FOOTER_PT_DATA_Y = data.get('FOOTER_PT_DATA_Y', FOOTER_PT_DATA_Y)
            FOOTER_PT_DATA_W = data.get('FOOTER_PT_DATA_W', FOOTER_PT_DATA_W)
            FOOTER_PT_DATA_H = data.get('FOOTER_PT_DATA_H', FOOTER_PT_DATA_H)

            START_ROW = data.get('START_ROW', START_ROW)
            # DATA_END_ROW = data.get('DATA_END_ROW', DATA_END_ROW)  # 코드 기본값 사용
            PRINT_END_ROW = data.get('PRINT_END_ROW', PRINT_END_ROW)
            
            print("✅ 저장된 설정을 불러왔습니다.")
        except Exception as e:
            print(f"⚠️ 설정 불러오기 실패 (기본값 사용): {e}")

def save_settings():
    data = {
        'SEOUL_COVER_ANCHOR': SEOUL_COVER_ANCHOR, 'SEOUL_COVER_X': SEOUL_COVER_X, 'SEOUL_COVER_Y': SEOUL_COVER_Y, 'SEOUL_COVER_W': SEOUL_COVER_W, 'SEOUL_COVER_H': SEOUL_COVER_H,
        'SITCO_COVER_ANCHOR': SITCO_COVER_ANCHOR, 'SITCO_COVER_X': SITCO_COVER_X, 'SITCO_COVER_Y': SITCO_COVER_Y, 'SITCO_COVER_W': SITCO_COVER_W, 'SITCO_COVER_H': SITCO_COVER_H,
        'FOOTER_COVER_ANCHOR': FOOTER_COVER_ANCHOR, 'FOOTER_COVER_X': FOOTER_COVER_X, 'FOOTER_COVER_Y': FOOTER_COVER_Y, 'FOOTER_COVER_W': FOOTER_COVER_W, 'FOOTER_COVER_H': FOOTER_COVER_H,
        'FOOTER_PT_COVER_ANCHOR': FOOTER_PT_COVER_ANCHOR, 'FOOTER_PT_COVER_X': FOOTER_PT_COVER_X, 'FOOTER_PT_COVER_Y': FOOTER_PT_COVER_Y, 'FOOTER_PT_COVER_W': FOOTER_PT_COVER_W, 'FOOTER_PT_COVER_H': FOOTER_PT_COVER_H,
        
        'SEOUL_DATA_ANCHOR': SEOUL_DATA_ANCHOR, 'SEOUL_DATA_X': SEOUL_DATA_X, 'SEOUL_DATA_Y': SEOUL_DATA_Y, 'SEOUL_DATA_W': SEOUL_DATA_W, 'SEOUL_DATA_H': SEOUL_DATA_H,
        'SITCO_DATA_ANCHOR': SITCO_DATA_ANCHOR, 'SITCO_DATA_X': SITCO_DATA_X, 'SITCO_DATA_Y': SITCO_DATA_Y, 'SITCO_DATA_W': SITCO_DATA_W, 'SITCO_DATA_H': SITCO_DATA_H,
        'FOOTER_DATA_ANCHOR': FOOTER_DATA_ANCHOR, 'FOOTER_DATA_X': FOOTER_DATA_X, 'FOOTER_DATA_Y': FOOTER_DATA_Y, 'FOOTER_DATA_W': FOOTER_DATA_W, 'FOOTER_DATA_H': FOOTER_DATA_H,
        'FOOTER_PT_DATA_ANCHOR': FOOTER_PT_DATA_ANCHOR, 'FOOTER_PT_DATA_X': FOOTER_PT_DATA_X, 'FOOTER_PT_DATA_Y': FOOTER_PT_DATA_Y, 'FOOTER_PT_DATA_W': FOOTER_PT_DATA_W, 'FOOTER_PT_DATA_H': FOOTER_PT_DATA_H,
        
        'START_ROW': START_ROW, 
        'DATA_END_ROW': DATA_END_ROW, 
        'PRINT_END_ROW': PRINT_END_ROW
    }
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4)
        print("✅ 설정이 저장되었습니다.")
    except Exception as e:
        print(f"⚠️ 설정 저장 실패: {e}")

# ========================================================
# [UI 함수] 설정 팝업
# ========================================================
def select_folder_ui(title):
    root = Tk(); root.withdraw(); root.attributes('-topmost', True) 
    dir_path = filedialog.askdirectory(title=title, initialdir=folder_path)
    root.destroy(); return dir_path

def select_file_ui(title, filetypes):
    root = Tk(); root.withdraw(); root.attributes('-topmost', True) 
    file_path = filedialog.askopenfilename(title=title, initialdir=folder_path, filetypes=filetypes)
    root.destroy(); return file_path

def open_position_settings_ui():
    global SEOUL_COVER_ANCHOR, SEOUL_COVER_X, SEOUL_COVER_Y, SEOUL_COVER_W, SEOUL_COVER_H
    global SITCO_COVER_ANCHOR, SITCO_COVER_X, SITCO_COVER_Y, SITCO_COVER_W, SITCO_COVER_H
    global FOOTER_COVER_ANCHOR, FOOTER_COVER_X, FOOTER_COVER_Y, FOOTER_COVER_W, FOOTER_COVER_H
    global FOOTER_PT_COVER_ANCHOR, FOOTER_PT_COVER_X, FOOTER_PT_COVER_Y, FOOTER_PT_COVER_W, FOOTER_PT_COVER_H
    
    global SEOUL_DATA_ANCHOR, SEOUL_DATA_X, SEOUL_DATA_Y, SEOUL_DATA_W, SEOUL_DATA_H
    global SITCO_DATA_ANCHOR, SITCO_DATA_X, SITCO_DATA_Y, SITCO_DATA_W, SITCO_DATA_H
    global FOOTER_DATA_ANCHOR, FOOTER_DATA_X, FOOTER_DATA_Y, FOOTER_DATA_W, FOOTER_DATA_H
    global FOOTER_PT_DATA_ANCHOR, FOOTER_PT_DATA_X, FOOTER_PT_DATA_Y, FOOTER_PT_DATA_W, FOOTER_PT_DATA_H
    
    global START_ROW, DATA_END_ROW, PRINT_END_ROW

    win = Tk(); win.title("세부 설정 (기준 셀/위치/크기)"); win.geometry("600x750"); win.attributes('-topmost', True)
    
    tabControl = ttk.Notebook(win)
    tab1 = ttk.Frame(tabControl); tab2 = ttk.Frame(tabControl); tab3 = ttk.Frame(tabControl)
    tabControl.add(tab1, text='[갑지] Cover'); tabControl.add(tab2, text='[을지] Data'); tabControl.add(tab3, text='[행] Rows')
    tabControl.pack(expand=1, fill="both")
    
    entries = {}
    
    def create_section(parent, row_start, title, defaults, key_prefix):
        Label(parent, text=title, font=("bold", 11)).grid(row=row_start, column=0, columnspan=6, pady=(15, 5), sticky="w")
        Label(parent, text="기준 셀:", fg="blue").grid(row=row_start+1, column=0, sticky='e')
        var_a = StringVar(value=str(defaults[0])); Entry(parent, textvariable=var_a, width=8, bg="#f0f8ff").grid(row=row_start+1, column=1, padx=2); entries[f"{key_prefix}_anchor"] = var_a
        Label(parent, text="(예: E4)").grid(row=row_start+1, column=2, sticky='w')

        Label(parent, text="가로(X):").grid(row=row_start+2, column=0, sticky='e')
        var_x = StringVar(value=str(defaults[1])); Entry(parent, textvariable=var_x, width=8).grid(row=row_start+2, column=1, padx=2); entries[f"{key_prefix}_x"] = var_x
        Label(parent, text="세로(Y):").grid(row=row_start+2, column=2, sticky='e')
        var_y = StringVar(value=str(defaults[2])); Entry(parent, textvariable=var_y, width=8).grid(row=row_start+2, column=3, padx=2); entries[f"{key_prefix}_y"] = var_y
        
        Label(parent, text="너비(W):").grid(row=row_start+3, column=0, sticky='e')
        var_w = StringVar(value=str(defaults[3])); Entry(parent, textvariable=var_w, width=8).grid(row=row_start+3, column=1, padx=2); entries[f"{key_prefix}_w"] = var_w
        Label(parent, text="높이(H):").grid(row=row_start+3, column=2, sticky='e')
        var_h = StringVar(value=str(defaults[4])); Entry(parent, textvariable=var_h, width=8).grid(row=row_start+3, column=3, padx=2); entries[f"{key_prefix}_h"] = var_h

    # 갑지
    create_section(tab1, 0, "1. SITCO (왼쪽 상단)", (SITCO_COVER_ANCHOR, SITCO_COVER_X, SITCO_COVER_Y, SITCO_COVER_W, SITCO_COVER_H), 'cover_sitco')
    create_section(tab1, 4, "2. 서울검사 (오른쪽 상단)", (SEOUL_COVER_ANCHOR, SEOUL_COVER_X, SEOUL_COVER_Y, SEOUL_COVER_W, SEOUL_COVER_H), 'cover_seoul')
    create_section(tab1, 8, "3. 바닥글 오른쪽", (FOOTER_COVER_ANCHOR, FOOTER_COVER_X, FOOTER_COVER_Y, FOOTER_COVER_W, FOOTER_COVER_H), 'cover_footer')
    create_section(tab1, 12, "4. 바닥글 왼쪽", (FOOTER_PT_COVER_ANCHOR, FOOTER_PT_COVER_X, FOOTER_PT_COVER_Y, FOOTER_PT_COVER_W, FOOTER_PT_COVER_H), 'cover_pt')

    # 을지
    create_section(tab2, 0, "1. SITCO (왼쪽 상단)", (SITCO_DATA_ANCHOR, SITCO_DATA_X, SITCO_DATA_Y, SITCO_DATA_W, SITCO_DATA_H), 'data_sitco')
    create_section(tab2, 4, "2. 서울검사 (오른쪽 상단)", (SEOUL_DATA_ANCHOR, SEOUL_DATA_X, SEOUL_DATA_Y, SEOUL_DATA_W, SEOUL_DATA_H), 'data_seoul')
    create_section(tab2, 8, "3. 바닥글 오른쪽", (FOOTER_DATA_ANCHOR, FOOTER_DATA_X, FOOTER_DATA_Y, FOOTER_DATA_W, FOOTER_DATA_H), 'data_footer')
    create_section(tab2, 12, "4. 바닥글 왼쪽", (FOOTER_PT_DATA_ANCHOR, FOOTER_PT_DATA_X, FOOTER_PT_DATA_Y, FOOTER_PT_DATA_W, FOOTER_PT_DATA_H), 'data_pt')

    # 행 설정
    Label(tab3, text="데이터 행 범위 설정", font=("bold", 12)).pack(pady=20)
    frame_rows = Frame(tab3); frame_rows.pack()
    Label(frame_rows, text="시작 행 (Start):").grid(row=0, column=0, padx=5, pady=5)
    var_s = StringVar(value=str(START_ROW)); Entry(frame_rows, textvariable=var_s, width=10).grid(row=0, column=1, padx=5, pady=5); entries['start_row'] = var_s
    Label(frame_rows, text="데이터 종료 행 (Data End):").grid(row=1, column=0, padx=5, pady=5)
    var_d_e = StringVar(value=str(DATA_END_ROW)); Entry(frame_rows, textvariable=var_d_e, width=10).grid(row=1, column=1, padx=5, pady=5); entries['data_end_row'] = var_d_e
    Label(frame_rows, text="인쇄 영역 종료 행 (Print End):").grid(row=2, column=0, padx=5, pady=5)
    var_p_e = StringVar(value=str(PRINT_END_ROW)); Entry(frame_rows, textvariable=var_p_e, width=10).grid(row=2, column=1, padx=5, pady=5); entries['print_end_row'] = var_p_e
    Label(tab3, text="* 데이터 종료 행: 데이터가 채워지는 마지막 줄 (예: 41)\n* 인쇄 영역 종료 행: 인쇄가 끝나는 줄 (예: 43)", fg="gray", justify="left").pack(pady=10)

    def on_apply():
        global SEOUL_COVER_ANCHOR, SEOUL_COVER_X, SEOUL_COVER_Y, SEOUL_COVER_W, SEOUL_COVER_H
        global SITCO_COVER_ANCHOR, SITCO_COVER_X, SITCO_COVER_Y, SITCO_COVER_W, SITCO_COVER_H
        global FOOTER_COVER_ANCHOR, FOOTER_COVER_X, FOOTER_COVER_Y, FOOTER_COVER_W, FOOTER_COVER_H
        global FOOTER_PT_COVER_ANCHOR, FOOTER_PT_COVER_X, FOOTER_PT_COVER_Y, FOOTER_PT_COVER_W, FOOTER_PT_COVER_H
        
        global SEOUL_DATA_ANCHOR, SEOUL_DATA_X, SEOUL_DATA_Y, SEOUL_DATA_W, SEOUL_DATA_H
        global SITCO_DATA_ANCHOR, SITCO_DATA_X, SITCO_DATA_Y, SITCO_DATA_W, SITCO_DATA_H
        global FOOTER_DATA_ANCHOR, FOOTER_DATA_X, FOOTER_DATA_Y, FOOTER_DATA_W, FOOTER_DATA_H
        global FOOTER_PT_DATA_ANCHOR, FOOTER_PT_DATA_X, FOOTER_PT_DATA_Y, FOOTER_PT_DATA_W, FOOTER_PT_DATA_H
        
        global START_ROW, DATA_END_ROW, PRINT_END_ROW
        
        try:
            SITCO_COVER_ANCHOR=entries['cover_sitco_anchor'].get(); SITCO_COVER_X=float(entries['cover_sitco_x'].get()); SITCO_COVER_Y=float(entries['cover_sitco_y'].get()); SITCO_COVER_W=float(entries['cover_sitco_w'].get()); SITCO_COVER_H=float(entries['cover_sitco_h'].get())
            SEOUL_COVER_ANCHOR=entries['cover_seoul_anchor'].get(); SEOUL_COVER_X=float(entries['cover_seoul_x'].get()); SEOUL_COVER_Y=float(entries['cover_seoul_y'].get()); SEOUL_COVER_W=float(entries['cover_seoul_w'].get()); SEOUL_COVER_H=float(entries['cover_seoul_h'].get())
            FOOTER_COVER_ANCHOR=entries['cover_footer_anchor'].get(); FOOTER_COVER_X=float(entries['cover_footer_x'].get()); FOOTER_COVER_Y=float(entries['cover_footer_y'].get()); FOOTER_COVER_W=float(entries['cover_footer_w'].get()); FOOTER_COVER_H=float(entries['cover_footer_h'].get())
            FOOTER_PT_COVER_ANCHOR=entries['cover_pt_anchor'].get(); FOOTER_PT_COVER_X=float(entries['cover_pt_x'].get()); FOOTER_PT_COVER_Y=float(entries['cover_pt_y'].get()); FOOTER_PT_COVER_W=float(entries['cover_pt_w'].get()); FOOTER_PT_COVER_H=float(entries['cover_pt_h'].get())

            SITCO_DATA_ANCHOR=entries['data_sitco_anchor'].get(); SITCO_DATA_X=float(entries['data_sitco_x'].get()); SITCO_DATA_Y=float(entries['data_sitco_y'].get()); SITCO_DATA_W=float(entries['data_sitco_w'].get()); SITCO_DATA_H=float(entries['data_sitco_h'].get())
            SEOUL_DATA_ANCHOR=entries['data_seoul_anchor'].get(); SEOUL_DATA_X=float(entries['data_seoul_x'].get()); SEOUL_DATA_Y=float(entries['data_seoul_y'].get()); SEOUL_DATA_W=float(entries['data_seoul_w'].get()); SEOUL_DATA_H=float(entries['data_seoul_h'].get())
            FOOTER_DATA_ANCHOR=entries['data_footer_anchor'].get(); FOOTER_DATA_X=float(entries['data_footer_x'].get()); FOOTER_DATA_Y=float(entries['data_footer_y'].get()); FOOTER_DATA_W=float(entries['data_footer_w'].get()); FOOTER_DATA_H=float(entries['data_footer_h'].get())
            FOOTER_PT_DATA_ANCHOR=entries['data_pt_anchor'].get(); FOOTER_PT_DATA_X=float(entries['data_pt_x'].get()); FOOTER_PT_DATA_Y=float(entries['data_pt_y'].get()); FOOTER_PT_DATA_W=float(entries['data_pt_w'].get()); FOOTER_PT_DATA_H=float(entries['data_pt_h'].get())

            START_ROW = int(entries['start_row'].get())
            DATA_END_ROW = int(entries['data_end_row'].get())
            PRINT_END_ROW = int(entries['print_end_row'].get())
            
            save_settings()
            win.destroy()
        except ValueError: messagebox.showerror("오류", "숫자는 정확히 입력해주세요.")

    Button(win, text="설정 적용 및 시작 (Start)", command=on_apply, bg="#d0e8f2", font=("bold", 12), height=2, width=30).pack(pady=20)
    win.mainloop()

# ========================================================
# [함수] 로고 찾기 & 배치
# ========================================================
def find_image_smart(keyword, exclude_keyword=None):
    candidates = glob.glob(os.path.join(logo_folder_path, "*.*"))
    valid_extensions = ['.PNG', '.JPG', '.JPEG', '.BMP', '.GIF']
    for path in candidates:
        fname = os.path.basename(path).upper(); ext = os.path.splitext(path)[1].upper()
        if ext not in valid_extensions: continue 
        if keyword.upper() in fname:
            if exclude_keyword and exclude_keyword.upper() in fname: continue
            return path 
    return None

def place_image_freely(ws, img_path, anchor_cell_str, w, h, x_offset, y_offset):
    try:
        if not img_path or not os.path.exists(img_path): return
        
        original = PILImage.open(img_path).convert("RGBA")
        resized = original.resize((int(w), int(h)), PILImage.Resampling.LANCZOS)
        temp_name = f"temp_{os.path.basename(img_path)}"
        # 시스템 임시 폴더 사용 (어느 위치에서 실행해도 작동)
        temp_full_path = os.path.join(tempfile.gettempdir(), temp_name)
        resized.save(temp_full_path)
        
        img = XLImage(temp_full_path); img.width = w; img.height = h
        col_str, row_num = coordinate_from_string(anchor_cell_str)
        col_idx = column_index_from_string(col_str) - 1; row_idx = row_num - 1 
        
        emu_x = int(x_offset * 9525); emu_y = int(y_offset * 9525)
        emu_w = int(w * 9525); emu_h = int(h * 9525)
        
        marker = AnchorMarker(col=col_idx, colOff=emu_x, row=row_idx, rowOff=emu_y)
        size = XDRPositiveSize2D(cx=emu_w, cy=emu_h)
        img.anchor = OneCellAnchor(_from=marker, ext=size)
        ws.add_image(img)
    except Exception as e: print(f"⚠️ 이미지 배치 에러: {e}")

def add_logos_to_sheet(ws, is_cover=False):
    if not logo_folder_path: return 
    try: ws._images = [] 
    except: pass
    
    if is_cover:
        s_sitco = (SITCO_COVER_ANCHOR, SITCO_COVER_W, SITCO_COVER_H, SITCO_COVER_X, SITCO_COVER_Y)
        s_seoul = (SEOUL_COVER_ANCHOR, SEOUL_COVER_W, SEOUL_COVER_H, SEOUL_COVER_X, SEOUL_COVER_Y)
        s_footer = (FOOTER_COVER_ANCHOR, FOOTER_COVER_W, FOOTER_COVER_H, FOOTER_COVER_X, FOOTER_COVER_Y)
        s_pt = (FOOTER_PT_COVER_ANCHOR, FOOTER_PT_COVER_W, FOOTER_PT_COVER_H, FOOTER_PT_COVER_X, FOOTER_PT_COVER_Y)
    else:
        s_sitco = (SITCO_DATA_ANCHOR, SITCO_DATA_W, SITCO_DATA_H, SITCO_DATA_X, SITCO_DATA_Y)
        s_seoul = (SEOUL_DATA_ANCHOR, SEOUL_DATA_W, SEOUL_DATA_H, SEOUL_DATA_X, SEOUL_DATA_Y)
        s_footer = (FOOTER_DATA_ANCHOR, FOOTER_DATA_W, FOOTER_DATA_H, FOOTER_DATA_X, FOOTER_DATA_Y)
        s_pt = (FOOTER_PT_DATA_ANCHOR, FOOTER_PT_DATA_W, FOOTER_PT_DATA_H, FOOTER_PT_DATA_X, FOOTER_PT_DATA_Y)

    path = find_image_smart("SITCO"); 
    if path: place_image_freely(ws, path, *s_sitco)
    path = find_image_smart("서울검사"); 
    if path: place_image_freely(ws, path, *s_seoul)
    path = find_image_smart("바닥글", exclude_keyword="PMI-1"); 
    if not path: path = find_image_smart("PMI", exclude_keyword="PMI-1")
    if path: place_image_freely(ws, path, *s_footer)
    
    path_left = None
    if is_cover:
        path_left = find_image_smart("PMI갑")
    if not path_left: path_left = find_image_smart("PMI-1")
    if not path_left: path_left = find_image_smart("PT")
    if path_left: place_image_freely(ws, path_left, *s_pt)

def cleanup_temp_files():
    # 시스템 임시 폴더에서 임시 파일 정리
    temp_dir = tempfile.gettempdir()
    temp_files = glob.glob(os.path.join(temp_dir, "temp_*.png"))
    for f in temp_files:
        try: os.remove(f)
        except: pass

# ========================================================
# [함수] 데이터 처리 및 엑셀 유틸
# ========================================================
def to_float(val):
    if pd.isna(val): return 0.0
    s = str(val).upper().replace("%", "").strip()
    if "<" in s or "ND" in s or s == "": return 0.0
    try: return float(s)
    except ValueError: return 0.0

def guess_grade(mo_val, original_grade_text):
    if original_grade_text and str(original_grade_text).strip() != "": return str(original_grade_text).strip()
    try: return "SS316" if float(mo_val) >= 1.5 else "SS304"
    except: return ""

def find_col(df, keywords, exclude_keywords=None):
    if exclude_keywords is None: exclude_keywords = []
    for col in df.columns:
        c_str = str(col).upper().strip()
        if any(ex in c_str for ex in exclude_keywords): continue
        if any(k in c_str for k in keywords): return col
    return None

def force_copy_footer(source_ws, target_ws):
    try:
        if source_ws.oddFooter: target_ws.oddFooter.left.text = source_ws.oddFooter.left.text
    except: pass 

def force_print_settings(source_ws, target_ws):
    try:
        target_ws.print_area = f'A1:R{PRINT_END_ROW}'  # R열까지
        
        # 기존 페이지 나누기 모두 제거
        target_ws.row_breaks.brk = []
        target_ws.col_breaks.brk = []
        
        # ★ 페이지 설정
        target_ws.page_setup.paperSize = 9  # A4
        target_ws.page_setup.orientation = 'landscape'  # 가로
        target_ws.page_setup.fitToHeight = 1  # 세로 1페이지
        target_ws.page_setup.fitToWidth = 1   # 가로 1페이지
        target_ws.page_setup.scale = None     # 배율 제거 (fitTo와 충돌 방지)
        
        # ★ fitToPage 활성화 (sheet_properties 사용)
        target_ws.sheet_properties.pageSetUpPr.fitToPage = True
        
        # ★ 인쇄 옵션 (가운데 정렬)
        target_ws.print_options.horizontalCentered = True
        target_ws.print_options.verticalCentered = False
            
        # ★ 여백 고정값 (좁은 여백)
        target_ws.page_margins.left = 0.3
        target_ws.page_margins.right = 0.3
        target_ws.page_margins.top = 0.4
        target_ws.page_margins.bottom = 0.4
        target_ws.page_margins.header = 0.2
        target_ws.page_margins.footer = 0.2
    except Exception as e: 
        print(f"⚠️ 인쇄 설정 오류: {e}")

def set_eulji_headers(ws):
    """을지 시트에 H20~P20 헤더 설정 + Q20 ~ Q35 재질 목록 (VBA 연동용 단일 목록)"""
    headers = ["NI", "CR", "MO", "NI", "CR", "MO", "NI", "CR", "MO"]
    data_font = Font(size=9)
    for i, val in enumerate(headers):
        col = 8 + i  # H=8, I=9, ... P=16
        cell = ws.cell(row=20, column=col)
        cell.value = val
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = data_font
    
    # [1] 데이터 유효성 검사 설정 (단일 항목 목록)
    # VBA 매크로가 이 목록에서 선택된 값을 기존 값 뒤에 추가(stack)할 것입니다.
    materials = "SS304,SS304L,SS316,SS316L,SS321,SS347,SS410,SS430,DUPLEX,MONEL,INCONEL,ER316L,WP316,WP316L,TP316,TP316L,F316L,A182-F316L,A312-TP316L"
    dv = DataValidation(type="list", formula1=f'"{materials}"', allow_blank=True)
    dv.error ='목록에 있는 항목 중 하나를 선택하십시오.'
    dv.errorTitle = '잘못된 입력'
    dv.prompt = '재질을 선택하세요. (최대 3개까지 추가 선택 가능)'
    dv.promptTitle = '재질 목록'
    ws.add_data_validation(dv)
    
    # Q20 ~ Q35(DATA_END_ROW) 범위에 유효성 검사 및 스타일 추가
    for r in range(START_ROW, DATA_END_ROW + 1):
        target_cell = ws.cell(row=r, column=17) # Q열
        dv.add(target_cell)
        target_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        target_cell.font = data_font

    # R20 셀 내용 지우기
    ws["R20"].value = None

def prepare_next_sheet(wb, source_sheet_idx, page_num):
    source_sheet = wb.worksheets[source_sheet_idx]; new_sheet = wb.copy_worksheet(source_sheet) 
    base_title = source_sheet.title.split('_')[0]; new_sheet.title = f"{base_title[:20]}_{page_num:03d}"
    new_sheet.sheet_view.tabSelected = False
    
    # 페이지 번호 설정 추가 (을지 시트 O3)
    try: new_sheet["O3"] = f"{page_num}"
    except: pass
    
    force_copy_footer(source_sheet, new_sheet)
    force_print_settings(source_sheet, new_sheet)
    add_logos_to_sheet(new_sheet, is_cover=False)
    
    # 데이터 영역 초기화 (7열~17열) 및 행 높이/글자크기 설정
    data_font = Font(size=9)
    for r in range(START_ROW, DATA_END_ROW + 1):
        new_sheet.row_dimensions[r].height = 40  # 3줄 텍스트 수용 (글자크기 9 고려)
        for c in range(7, 18):  # G~Q열 초기화 및 스타일 적용
            cell = new_sheet.cell(row=r, column=c)
            cell.font = data_font
            if not isinstance(cell, MergedCell): cell.value = None
            else:
                for rng in new_sheet.merged_cells.ranges:
                    if cell.coordinate in rng: new_sheet.cell(row=rng.min_row, column=rng.min_col).value = None; break
    
    # 38~41행 삭제 (4행 삭제)
    new_sheet.delete_rows(38, 4)
    
    return new_sheet

# ========================================================
# [핵심] 엑셀 데이터 채우기 (좌우 분할 & 줄바꿈 모드)
# ========================================================
def fill_excel_pagination(wb, data_list):
    # [설정] 스타일 정의
    wrap_style = Alignment(wrap_text=True, horizontal='center', vertical='center', shrink_to_fit=True)
    center_style = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)

    current_sheet_idx = 1 if len(wb.worksheets) >= 2 else 0
    current_page_num = 1
    ws = wb.worksheets[current_sheet_idx]
    
    # 갑지/을지 여부 먼저 확인
    is_cover = (current_sheet_idx == 0)
    
    # 시트명/페이지번호/로고 설정
    ws.title = f"{ws.title[:20]}_{current_page_num:03d}"
    # 갑지 시트는 N3에, 을지 시트는 Q3에 페이지 번호 표시
    if is_cover:
        try: ws["N3"] = f"Page {current_page_num} of {len(wb.worksheets)}."
        except: pass 
    else:
        try: ws["Q3"] = f"Page {current_page_num} of {len(wb.worksheets)}."
        except: pass 
    
    add_logos_to_sheet(ws, is_cover=is_cover)
    force_print_settings(ws, ws)
    
    # 을지 시트인 경우 헤더 설정 및 행 높이/글자크기 조정
    if not is_cover:
        set_eulji_headers(ws)
        data_font = Font(size=9)
        for r in range(START_ROW, DATA_END_ROW + 1):
            ws.row_dimensions[r].height = 40  # 3줄 텍스트 수용 (글자크기 9 고려)
            for c in range(7, 18): # G~Q열
                ws.cell(row=r, column=c).font = data_font
        # 38~41행 삭제 (4행 삭제)
        ws.delete_rows(38, 4)
    
    # ========================================================
    # [컨럼 매핑] 3개 데이터씩 - NI, CR, MO 순서 (탄소 제외)
    # ========================================================
    col_no = 7      # G열

    # 첫 번째 데이터 - H, I, J
    map_1 = {'Ni': 8, 'Cr': 9, 'Mo': 10}
    
    # 두 번째 데이터 - K, L, M
    map_2 = {'Ni': 11, 'Cr': 12, 'Mo': 13}
    
    # 세 번째 데이터 - N, O, P
    map_3 = {'Ni': 14, 'Cr': 15, 'Mo': 16}

    current_row = START_ROW 
    
    for i in range(0, len(data_list), 3):  # 3개씩 처리
        if current_row > DATA_END_ROW:
            current_page_num += 1
            ws = prepare_next_sheet(wb, current_sheet_idx, current_page_num)
            try: ws["Q3"] = f"{current_page_num}"
            except: pass
            current_sheet_idx += 1
            current_row = START_ROW 
        
        item1 = data_list[i]
        item2 = data_list[i+1] if (i + 1) < len(data_list) else None
        item3 = data_list[i+2] if (i + 2) < len(data_list) else None

        # (1) G열: No (최대 3개 줄바꿈)
        cell_no = ws.cell(row=current_row, column=col_no)
        no_list = [str(item1['No'])]
        if item2: no_list.append(str(item2['No']))
        if item3: no_list.append(str(item3['No']))
        
        if len(no_list) > 1:
            cell_no.value = "\n".join(no_list)
            cell_no.alignment = wrap_style
        else:
            cell_no.value = item1['No']
            cell_no.alignment = center_style

        # (2) H~K열: 첫 번째 데이터
        for key, col_idx in map_1.items():
            val = item1.get(key, 0.0)
            if val == 0 or val == 0.0: val = ""
            elif isinstance(val, (int, float)): val = round(val, 1)
            c = ws.cell(row=current_row, column=col_idx)
            c.value = val
            c.alignment = center_style
            if val != "": c.number_format = '0.0'

        # (3) L~O열: 두 번째 데이터
        if item2:
            for key, col_idx in map_2.items():
                val = item2.get(key, 0.0)
                if val == 0 or val == 0.0: val = ""
                elif isinstance(val, (int, float)): val = round(val, 1)
                c = ws.cell(row=current_row, column=col_idx)
                c.value = val
                c.alignment = center_style
                if val != "": c.number_format = '0.0'

        # (4) P~S열: 세 번째 데이터
        if item3:
            for key, col_idx in map_3.items():
                val = item3.get(key, 0.0)
                if val == 0 or val == 0.0: val = ""
                elif isinstance(val, (int, float)): val = round(val, 1)
                c = ws.cell(row=current_row, column=col_idx)
                c.value = val
                c.alignment = center_style
                if val != "": c.number_format = '0.0'

        # (5) Q열: Grade (최대 3개 줄바꿈)
        cell_grade = ws.cell(row=current_row, column=17) # Q열
        grade_list = [str(item1.get('Grade', ''))]
        if item2: grade_list.append(str(item2.get('Grade', '')))
        if item3: grade_list.append(str(item3.get('Grade', '')))
        
        cell_grade.value = "\n".join(grade_list)
        cell_grade.alignment = wrap_style

        current_row += 1

    # 모든 시트에 "Page X of Y." 형식으로 페이지 번호 업데이트
    total_pages = len(wb.worksheets)
    page_style = Alignment(horizontal='distributed', vertical='center')
    for idx, sheet in enumerate(wb.worksheets):
        page_num = idx + 1
        if idx == 0:  # 갑시트
            try: 
                sheet["N3"] = f"Page {page_num} of {total_pages}."
                sheet["N3"].alignment = page_style
            except: pass
        else:  # 을시트
            try: 
                sheet["O3"] = f"Page {page_num} of {total_pages}."
                sheet["O3"].alignment = page_style
            except: pass

    print(f"   ✅ 데이터 입력 완료 (3개씩 분할 / 빈칸 처리 적용됨)")

# ========================================================
# [메인] 실행
# ========================================================
try:
    print("\n" + "="*30)
    print("[1] 로고 파일 확인 및 설정")
    logo_sitco = find_image_smart("SITCO"); logo_seoul = find_image_smart("서울검사")
    if not logo_sitco or not logo_seoul:
        print("👉 'SITCO'와 '서울검사' 로고가 들어있는 폴더를 선택해주세요.")
        selected_dir = select_folder_ui("로고 이미지 파일이 있는 폴더를 선택하세요")
        if selected_dir: logo_folder_path = selected_dir
    load_settings(); open_position_settings_ui()

    print("\n[2] 데이터 파일(RFI) 선택"); 
    target_file = select_file_ui("분석할 원본 데이터 선택 (RFI)", [("Excel files", "*.xls;*.xlsx;*.xlsm")])
    if not target_file: sys.exit()
    print(f"📂 선택됨: {os.path.basename(target_file)}")

    print("\n[3] 양식 파일(SIT-PMI) 선택"); 
    template_path = select_file_ui("SIT-PMI 양식 파일 선택", [("Excel files", "*.xlsx;*.xlsm")])
    if not template_path: sys.exit()
    print(f"📂 선택됨: {os.path.basename(template_path)}")

    print("\n[4] 순번 선택 (선택 사항)")
    root = Tk(); root.withdraw(); root.attributes('-topmost', True)
    target_input = simpledialog.askstring("순번 선택", "특정 순번(NO)만 하려면 입력\n(전체 처리는 취소 또는 빈칸):", parent=root)
    root.destroy()
    target_no_list = []
    if target_input: target_no_list = [x.strip() for x in target_input.replace(',', ' ').split() if x.strip()]

    all_extracted_data = []
    xls = pd.ExcelFile(target_file)
    for sheet_name in xls.sheet_names:
        print(f"   [Processing] {sheet_name}...")
        try: temp_df = pd.read_excel(target_file, sheet_name=sheet_name, header=None, nrows=50)
        except: continue
        header_idx = None
        for i, row in temp_df.iterrows():
            row_str = str(row.values).upper()
            if ("CR" in row_str and "NI" in row_str) or ("CHROMIUM" in row_str): header_idx = i; break
        
        if header_idx is None:
            print(f"     -> ⚠️ 헤더 자동 검색 실패. 수동 입력 필요.")
            root = Tk(); root.withdraw(); root.attributes('-topmost', True)
            user_input = simpledialog.askstring("헤더 행 지정", f"'{sheet_name}' 시트의 제목 줄 번호를 입력하세요:\n(건너뛰려면 취소)", parent=root)
            root.destroy()
            if user_input and user_input.isdigit(): header_idx = int(user_input)
            else: continue

        df = pd.read_excel(target_file, sheet_name=sheet_name, header=header_idx)
        col_cr = find_col(df, ["CR", "CHROMIUM"]); col_ni = find_col(df, ["NI", "NICKEL"])
        
        if not (col_cr and col_ni):
            print("\n⚠️ Cr/Ni 컬럼 인식 실패. 수동 지정 필요.")
            col_list = "\n".join([f"[{idx}] {col_name}" for idx, col_name in enumerate(df.columns)])
            root = Tk(); root.withdraw(); root.attributes('-topmost', True)
            try:
                idx_cr = simpledialog.askstring("컬럼 지정", f"컬럼 목록:\n{col_list}\n\nCr 컬럼 번호 입력:", parent=root)
                idx_ni = simpledialog.askstring("컬럼 지정", f"Ni 컬럼 번호 입력:", parent=root)
                idx_mo = simpledialog.askstring("컬럼 지정", f"Mo 컬럼 번호 입력 (선택사항):", parent=root)
                root.destroy()
                if idx_cr and idx_ni and idx_cr.isdigit() and idx_ni.isdigit():
                    col_cr = df.columns[int(idx_cr)]; col_ni = df.columns[int(idx_ni)]
                    col_mo = df.columns[int(idx_mo)] if idx_mo and idx_mo.isdigit() else None
                else: continue
            except: root.destroy(); continue

        col_mo = col_mo if 'col_mo' in locals() and col_mo else find_col(df, ["MO", "MOLYBDENUM"])
        col_mn = find_col(df, ["MN", "MANGANESE"])
        col_no = find_col(df, ["NO.", "NO", "SEQ", "NUM", "POS", "ITEM"]) 
        col_grade_orig = find_col(df, ["GRADE", "MATERIAL", "SPEC", "TYPE"])

        df['val_Cr'] = df[col_cr].apply(to_float); df['val_Ni'] = df[col_ni].apply(to_float)
        df['val_Mo'] = df[col_mo].apply(to_float) if col_mo else 0.0
        df['val_Mn'] = df[col_mn].apply(to_float) if col_mn else 0.0
        if col_no: df['val_No'] = df[col_no].astype(str).str.strip()
        else: df['val_No'] = range(1, len(df) + 1)
        df['val_Grade_Orig'] = df[col_grade_orig] if col_grade_orig else ""

        count_in_sheet = 0
        for _, row in df.iterrows():
            current_no = str(row['val_No'])
            if target_no_list and current_no not in target_no_list: continue 
            if row['val_Cr'] > 0 or (str(row['val_No']) != "" and str(row['val_No']) != "nan"):
                all_extracted_data.append({
                    'No': row['val_No'], 'Cr': row['val_Cr'], 'Ni': row['val_Ni'], 
                    'Mo': row['val_Mo'], 'Mn': row['val_Mn'], 'Grade': guess_grade(row['val_Mo'], row['val_Grade_Orig'])
                })
                count_in_sheet += 1
        print(f"     -> {count_in_sheet}개 데이터 확보")

    if not all_extracted_data: print("\n❌ 추출된 데이터가 없습니다."); sys.exit()

    # [추가] 사용자가 입력한 순번 리스트(target_no_list)의 순서에 맞춰 데이터 정렬
    if target_no_list:
        try:
            all_extracted_data.sort(key=lambda x: target_no_list.index(str(x['No'])) if str(x['No']) in target_no_list else 999999)
            print("   ✅ 입력하신 순번 순서대로 데이터 정렬 완료")
        except Exception as e:
            print(f"   ⚠️ 정렬 중 오류 발생 (원본 순서 유지): {e}")

    now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = os.path.splitext(os.path.basename(template_path))[0]
    # 양식 파일 확장자 유지 (.xlsm 인 경우 그대로 .xlsm으로 저장)
    ext = os.path.splitext(template_path)[1]
    new_filename = f"{base_name}_완료_{now_str}{ext}"
    # 출력 파일을 양식 파일과 같은 폴더에 저장
    output_folder = os.path.dirname(os.path.abspath(template_path))
    new_save_path = os.path.join(output_folder, new_filename)

    try:
        # VBA가 포함된 경우 keep_vba=True 설정
        wb = openpyxl.load_workbook(template_path, keep_vba=True)
        # 외부 참조 제거 옵션 추가
        if hasattr(wb, 'external_links'):
            try:
                wb.external_links.clear()
                print("   ✅ 외부 참조 링크 제거 완료")
            except:
                print("   ⚠️ 외부 참조 제거 실패 (계속 진행)")
        
        if len(wb.worksheets) >= 1: add_logos_to_sheet(wb.worksheets[0], is_cover=True)
        fill_excel_pagination(wb, all_extracted_data)
        for sheet in wb.worksheets: sheet.sheet_view.tabSelected = False
        wb.active = 0; wb.worksheets[0].sheet_view.tabSelected = True
        wb.save(new_save_path)
        print(f"\n💾 완료! 저장된 파일: {new_filename}"); print(f"📂 폴더: {logo_folder_path}")
    except PermissionError: print(f"\n🚫 [오류] 엑셀 파일이 열려있습니다. 닫고 다시 시도하세요.")
    except Exception as e: print(f"\n🚫 오류: {e}"); traceback.print_exc()

except Exception as e: traceback.print_exc(); print("오류 발생")
finally: 
    cleanup_temp_files()
    try:
        root = Tk(); root.withdraw(); root.attributes('-topmost', True)
        messagebox.showinfo("완료", "작업이 완료되었습니다.", parent=root)
        root.destroy()
    except:
        pass