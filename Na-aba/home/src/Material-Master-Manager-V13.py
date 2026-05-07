### VERSION: BUDGET_SYNC_FIXED_FINAL_V6_STRUCTURAL_OK ###
import mimetypes
import os
# [OPTIMIZATION] Prevent slow Windows registry scan for mimetypes
if os.name == 'nt':
    mimetypes.init(files=[])

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sys
import subprocess
import os
import time
import datetime
import json
import ctypes
import re
import traceback

def install_and_import(package, import_name=None):
    if import_name is None: import_name = package
    try:
        return __import__(import_name)
    except ImportError:
        try:
            print(f"Installing {package}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package, "--break-system-packages"])
            return __import__(import_name)
        except Exception:
            # Fallback for uv-managed systems
            try:
                subprocess.check_call(["uv", "pip", "install", "--system", "--break-system-packages", package])
                return __import__(import_name)
            except:
                pass
            sys.exit(1)

# Pre-import critical libraries
pd = install_and_import('pandas')
np = install_and_import('numpy')
install_and_import('openpyxl')
install_and_import('tkcalendar')
install_and_import('xlsxwriter')
install_and_import('pillow', 'PIL')
from PIL import Image, ImageTk

from daily_work_report_manager import DailyWorkReportManager
import daily_work_report_manager
print(f"DEBUG: daily_work_report_manager path: {daily_work_report_manager.__file__}")

# Pre-compiled regex for performance
NAN_PATTERN = re.compile(r'^nan(\.0+)?$|^none$|^null$|^0\.0+|-0\.0+$', re.IGNORECASE)
DOT_ZERO_PATTERN = re.compile(r'\.0$')
MARKER_PATTERN = re.compile(r'\(.*?\)\s*|익일')

# Auto-install additional dependencies
from tkcalendar import DateEntry, Calendar

# [FIX] Patch tkcalendar for stability in ko_KR locale / Python 3.14.
# 각 패치를 독립적으로 적용 - 한 패치 실패가 나머지를 무효화하지 않도록 분리.
import tkinter as _tk_ref

# 1) Calendar.__init__ : select_on_nav=False 기본값 + 헤더 너비 조정
try:
    _orig_cal_init = Calendar.__init__
    def _patched_cal_init(self, *args, **kwargs):
        if 'select_on_nav' not in kwargs:
            kwargs['select_on_nav'] = False
        _orig_cal_init(self, *args, **kwargs)
        def widen():
            if hasattr(self, '_header_month'):
                try: self._header_month.configure(width=25)
                except: pass
            if hasattr(self, '_header_year'):
                try: self._header_year.configure(width=10)
                except: pass
        self.after_idle(widen)
        self._widen_headers = widen
    Calendar.__init__ = _patched_cal_init
except Exception:
    pass

# 2) Calendar._display_calendar : 헤더 너비 재조정
try:
    _orig_cal_display = Calendar._display_calendar
    def _patched_cal_display(self, *args, **kwargs):
        _orig_cal_display(self, *args, **kwargs)
        if hasattr(self, '_widen_headers'):
            self.after_idle(self._widen_headers)
    Calendar._display_calendar = _patched_cal_display
except Exception:
    pass

# 3) DateEntry._show_calendar : tkcalendar < 1.6 버전 호환 (1.6.1에서는 없음 - 무시)
try:
    if hasattr(DateEntry, '_show_calendar'):
        _orig_de_show_cal = DateEntry._show_calendar
        def _patched_de_show_cal(self):
            try: self._last_known_date = self.get_date()
            except: self._last_known_date = None
            try: _orig_de_show_cal(self)
            except BaseException: pass
            if hasattr(self, '_calendar'):
                try: self._calendar.configure(select_on_nav=False)
                except: pass
        DateEntry._show_calendar = _patched_de_show_cal
except Exception:
    pass

# --- GLOBAL UTILITY FUNCTIONS ---
def normalize_id(val):
    """Robustly normalize IDs: handle NaN, trailing .0, and whitespace."""
    if pd.isna(val) or val == '' or str(val).lower() == 'nan': return ""
    s = str(val).strip()
    if s.endswith('.0'): s = s[:-2]
    return s

# 4) DateEntry.drop_down : tkcalendar 1.6.1+ 날짜박스 클릭 시 튕김 방지 (핵심 패치)
try:
    if hasattr(DateEntry, 'drop_down'):
        _orig_drop_down = DateEntry.drop_down
        def _patched_drop_down(self):
            try: self._last_known_date = self.get_date()
            except: self._last_known_date = None
            try:
                _orig_drop_down(self)
                if hasattr(self, '_calendar'):
                    try: self._calendar.configure(select_on_nav=False)
                    except: pass
            except BaseException:
                pass
        DateEntry.drop_down = _patched_drop_down
except Exception:
    pass

# 5) DateEntry._on_b1_press : 클릭 이벤트 처리 중 예외 방지
try:
    if hasattr(DateEntry, '_on_b1_press'):
        _orig_b1_press = DateEntry._on_b1_press
        def _patched_b1_press(self, event):
            try: _orig_b1_press(self, event)
            except BaseException: pass
        DateEntry._on_b1_press = _patched_b1_press
except Exception:
    pass

# 6) DateEntry._on_calendar_selection
try:
    if hasattr(DateEntry, '_on_calendar_selection'):
        _orig_de_on_sel = DateEntry._on_calendar_selection
        def _patched_de_on_sel(self, event):
            try:
                new_date = self._calendar.selection_get()
                if hasattr(self, '_last_known_date') and self._last_known_date == new_date:
                    pass
            except: pass
            _orig_de_on_sel(self, event)
        DateEntry._on_calendar_selection = _patched_de_on_sel
except Exception:
    pass

# 7) DateEntry._setup_style : 스타일 설정 중 update_idletasks 일시 차단
try:
    if hasattr(DateEntry, '_setup_style'):
        _orig_setup_style = DateEntry._setup_style
        def _patched_setup_style(self):
            _orig_update = self.update_idletasks
            self.update_idletasks = lambda: None
            try:
                _orig_setup_style(self)
            except BaseException:
                pass
            finally:
                self.update_idletasks = _orig_update
        DateEntry._setup_style = _patched_setup_style
except Exception:
    pass

# 8) DateEntry.update_idletasks : Python 3.14 BaseException 안전화 (클래스 레벨 영구 적용)
try:
    def _safe_de_update_idletasks(self):
        try:
            _tk_ref.Misc.update_idletasks(self)
        except BaseException:
            pass
    DateEntry.update_idletasks = _safe_de_update_idletasks
except Exception:
    pass

# 9) DateEntry._determine_downarrow_name : Configure/Map 이벤트 콜백 예외 방지
try:
    if hasattr(DateEntry, '_determine_downarrow_name'):
        _orig_det = DateEntry._determine_downarrow_name
        def _patched_det(self, event=None):
            try: _orig_det(self, event)
            except BaseException: pass
        DateEntry._determine_downarrow_name = _patched_det
except Exception:
    pass

# 10) Calendar._setup_style : ThemeChanged 시 예외 방지
try:
    if hasattr(Calendar, '_setup_style'):
        _orig_cal_setup_style = Calendar._setup_style
        def _patched_cal_setup_style(self, event=None):
            try: _orig_cal_setup_style(self, event)
            except BaseException: pass
        Calendar._setup_style = _patched_cal_setup_style
except Exception:
    pass

# 11) Calendar._prev_year / _next_year : 년도 이동
# [FIX] Exception 수준에서만 잡고, 날짜도 롤백.
try:
    if hasattr(Calendar, '_prev_year'):
        _orig_prev_year = Calendar._prev_year
        def _patched_prev_year(self):
            _saved = self._date
            try:
                _orig_prev_year(self)
            except Exception:
                self._date = _saved   # 실패 시 날짜 롤백
        Calendar._prev_year = _patched_prev_year
except Exception:
    pass

try:
    if hasattr(Calendar, '_next_year'):
        _orig_next_year = Calendar._next_year
        def _patched_next_year(self):
            _saved = self._date
            try:
                _orig_next_year(self)
            except Exception:
                self._date = _saved   # 실패 시 날짜 롤백
        Calendar._next_year = _patched_next_year
except Exception:
    pass

# 12) Calendar._prev_month / _next_month : 월 이동도 동일하게 안전화
try:
    if hasattr(Calendar, '_prev_month'):
        _orig_prev_month = Calendar._prev_month
        def _patched_prev_month(self):
            _saved = self._date
            try:
                _orig_prev_month(self)
            except Exception:
                self._date = _saved
        Calendar._prev_month = _patched_prev_month
except Exception:
    pass

try:
    if hasattr(Calendar, '_next_month'):
        _orig_next_month = Calendar._next_month
        def _patched_next_month(self):
            _saved = self._date
            try:
                _orig_next_month(self)
            except Exception:
                self._date = _saved
        Calendar._next_month = _patched_next_month
except Exception:
    pass

# 13) DateEntry._on_focus_out_cal : 년도/월 네비게이션 버튼 클릭 시 달력 닫힘 방지
# [근본 원인] 년도 버튼(_l_year/_r_year) 클릭 → Calendar에 <FocusOut> 발생
# → focus_get()이 버튼 위젯(DateEntry가 아닌 _top_cal 자식)을 반환
# → 기존 코드: "focus != self(DateEntry)" 이므로 else: _top_cal.withdraw() → 달력 닫힘
# → 사용자 입장: 년도 버튼 누를 때마다 달력이 꺼짐 ("화면 튕김")
# [수정] focus가 _top_cal 내부 위젯으로 이동한 경우 → 달력 유지
try:
    if hasattr(DateEntry, '_on_focus_out_cal'):
        _orig_on_focus_out_cal = DateEntry._on_focus_out_cal
        def _patched_on_focus_out_cal(self, event):
            fw = self.focus_get()
            if fw is not None:
                try:
                    # focus가 _top_cal 내부(년도·월 버튼 등)로 이동한 경우 → 유지
                    if str(fw).startswith(str(self._top_cal)):
                        return
                except Exception:
                    pass
            try:
                _orig_on_focus_out_cal(self, event)
            except BaseException:
                pass
        DateEntry._on_focus_out_cal = _patched_on_focus_out_cal
except Exception:
    pass

# --- Custom Draggable Messagebox Implementation ---
class DraggableMessagebox:
    """A custom draggable replacements for standard tkinter.messagebox"""
    @staticmethod
    def _show(type, title, message):
        root = tk._default_root
        if not root:
            # Fallback to standard if no root exists yet
            if type == "error": return messagebox.showerror_orig(title, message)
            if type == "warning": return messagebox.showwarning_orig(title, message)
            return messagebox.showinfo_orig(title, message)

        dialog = tk.Toplevel(root)
        dialog.overrideredirect(True) # Remove standard title bar for better drag control
        dialog.attributes("-topmost", True)
        
        # Style the custom dialog
        dialog.config(background="#f3f4f6", highlightthickness=1, highlightbackground="#d1d5db")

        # Custom Title Bar
        title_bar = tk.Frame(dialog, background="#ffffff", height=30, cursor="fleur")
        title_bar.pack(side="top", fill="x")
        
        title_lbl = tk.Label(title_bar, text=title, font=("Malgun Gothic", 9, "bold"), background="#ffffff", padx=10)
        title_lbl.pack(side="left")
        
        def close_dialog():
            dialog.grab_release()
            dialog.destroy()

        btn_close = tk.Label(title_bar, text="✕", font=("Malgun Gothic", 10), background="#ffffff", padx=10, cursor="hand2")
        btn_close.pack(side="right")
        btn_close.bind("<Button-1>", lambda e: close_dialog())
        btn_close.bind("<Enter>", lambda e: btn_close.config(background="#ef4444", foreground="white"))
        btn_close.bind("<Leave>", lambda e: btn_close.config(background="#ffffff", foreground="black"))

        # Disable main window
        dialog.transient(root)
        dialog.grab_set()

        # Dragging logic
        def start_drag(event):
            dialog._drag_start_x = event.x
            dialog._drag_start_y = event.y

        def do_drag(event):
            x = dialog.winfo_x() + event.x - dialog._drag_start_x
            y = dialog.winfo_y() + event.y - dialog._drag_start_y
            dialog.geometry(f"+{x}+{y}")

        title_bar.bind("<Button-1>", start_drag)
        title_bar.bind("<B1-Motion>", do_drag)
        title_lbl.bind("<Button-1>", start_drag)
        title_lbl.bind("<B1-Motion>", do_drag)

        # Content
        main_frame = tk.Frame(dialog, background="#f3f4f6", padx=20, pady=20)
        main_frame.pack(expand=True, fill='both')

        icon_char = "ℹ" if type == "info" else "⚠" if type == "warning" else "❌"
        icon_color = "#0078d7" if type == "info" else "#f59e0b" if type == "warning" else "#ef4444"
        
        lbl_icon = tk.Label(main_frame, text=icon_char, font=("Malgun Gothic", 24), fg=icon_color, background="#f3f4f6")
        lbl_icon.pack(side="left", anchor="n", padx=(0, 15))

        # Use a wider wraplength for longer messages
        lbl_msg = tk.Label(main_frame, text=message, font=("Malgun Gothic", 10), justify="left", wraplength=480, background="#f3f4f6")
        lbl_msg.pack(side="left", fill="both", expand=True)

        # Auto-size: compute proper width/height after all widgets are built
        def _auto_position():
            dialog.update_idletasks()
            req_w = max(440, dialog.winfo_reqwidth() + 20)
            req_h = max(200, dialog.winfo_reqheight() + 20)
            x = root.winfo_x() + (root.winfo_width() // 2) - (req_w // 2)
            y = root.winfo_y() + (root.winfo_height() // 2) - (req_h // 2)
            dialog.geometry(f"{req_w}x{req_h}+{x}+{y}")
        dialog.after(10, _auto_position)

        btn_frame = tk.Frame(dialog, background="#f3f4f6", pady=10)
        btn_frame.pack(side="bottom", fill='x')
        
        # Standard tk.Button for crisp rectangular (사각형) look
        btn_ok = tk.Button(btn_frame, text="확인", font=("Malgun Gothic", 10, "bold"), 
                           command=close_dialog, background="#ffffff", activebackground="#e5e7eb",
                           relief="raised", borderwidth=1, padx=30, pady=5, width=10)
        btn_ok.pack(pady=10)
        btn_ok.focus_set()
        
        dialog.bind("<Return>", lambda e: close_dialog())
        dialog.bind("<Escape>", lambda e: close_dialog())

        dialog.lift()
        dialog.focus_force()
        root.wait_window(dialog)

    @staticmethod
    def showerror(title, message): DraggableMessagebox._show("error", title, message)
    @staticmethod
    def showwarning(title, message): DraggableMessagebox._show("warning", title, message)
    @staticmethod
    def showinfo(title, message): DraggableMessagebox._show("info", title, message)

# Injection: Replace standard messagebox methods to enable draggable behavior globally
if not hasattr(messagebox, 'showerror_orig'):
    messagebox.showerror_orig = messagebox.showerror
    messagebox.showwarning_orig = messagebox.showwarning
    messagebox.showinfo_orig = messagebox.showinfo
    
    messagebox.showerror = DraggableMessagebox.showerror
    messagebox.showwarning = DraggableMessagebox.showwarning
    messagebox.showinfo = DraggableMessagebox.showinfo







class WorkerCompositeWidget(ttk.Frame):
    """
    Composite widget for Worker selection: [Name] with Autocomplete
    """
    def __init__(self, parent, enable_autocomplete=False, user_list=None, **kwargs):
        super().__init__(parent)
        
        # Worker Name selection
        name_width = kwargs.pop('width', 15)
        self.cb_name = ttk.Combobox(self, width=name_width, **kwargs)
        self.cb_name.pack(side='left', fill='x', expand=True)
        
    def get(self):
        """Return clean name"""
        return self.cb_name.get().strip()

    def set(self, value):
        """Set name, cleaning off any (Shift) prefixes if present"""
        if not value:
            self.cb_name.set("")
            return
            
        import re
        # Progressively migrate: if data still has (Shift) prefix, strip it for the name field
        match = re.match(r"\((주간|야간|휴일|주야간)\)\s*(.*)", str(value))
        if match:
            self.cb_name.set(match.group(2).strip())
        else:
            self.cb_name.set(str(value).strip())

    def bind(self, sequence=None, func=None, add=None):
        self.cb_name.bind(sequence, func, add)

    def current(self, newindex=None):
        return self.cb_name.current(newindex)
        
    def config(self, **kwargs):
        self.cb_name.config(**kwargs)

    def __setitem__(self, key, value):
        self.cb_name[key] = value

    def __getitem__(self, key):
        return self.cb_name[key]

class WorkerDataGroup(ttk.Frame):
    """
    Unified widget for a worker's record: [Name] [Shift] [WorkTime] [OT]
    """
    def __init__(self, parent, worker_index, users_list, time_list=None, enable_autocomplete=False, **kwargs):
        super().__init__(parent, padding=2) # Reduced padding for compact layout
        self.worker_index = worker_index
        
        # 1. Name selection (WorkerCompositeWidget now handles only name)
        self.composite = WorkerCompositeWidget(
            self, width=12, values=users_list, 
            enable_autocomplete=enable_autocomplete, 
            user_list=users_list
        )
        self.composite.pack(side='left', padx=(0, 2))
        self.cb_name = self.composite.cb_name
        
        # 2. Shift selection (Moved here from WorkerCompositeWidget)
        self.cb_shift = ttk.Combobox(self, values=["주간", "야간", "휴일", "주야간"], width=6, state="readonly")
        self.cb_shift.pack(side='left', padx=(1, 2))
        self.cb_shift.set("") # Default empty
        
        # 3. Work Time (Changed to Combobox for mouse selection)
        ttk.Label(self, text="시간:").pack(side='left', padx=(1, 0))
        self.ent_worktime = ttk.Combobox(self, width=16, values=time_list or [])
        self.ent_worktime.pack(side='left', padx=(0, 2))
        self.ent_worktime.set("") # Default empty
        
        # 4. OT
        ttk.Label(self, text="OT:").pack(side='left', padx=(1, 0))
        self.ent_ot = ttk.Entry(self, width=18)
        self.ent_ot.pack(side='left')

    def get_worker(self): return self.composite.get()
    def set_worker(self, val): self.composite.set(val)
    
    def get_time(self): 
        """Return combined string: (Shift) Time"""
        shift = self.cb_shift.get()
        time = self.ent_worktime.get().strip()
        if not time:
            return ""
        return f"({shift}) {time}"
        
    def set_time(self, val):
        """Parse string '(Shift) Time' and set widgets"""
        if not val:
            self.ent_worktime.set("")
            self.cb_shift.set("주간")
            return
            
        import re
        match = re.match(r"\((주간|야간|휴일|주야간)\)\s*(.*)", str(val))
        if match:
            self.cb_shift.set(match.group(1))
            self.ent_worktime.set(match.group(2).strip())
        else:
            # Fallback for old format (just time)
            self.cb_shift.set("주간")
            self.ent_worktime.set(str(val).strip())

    def get_ot(self): return self.ent_ot.get()
    def set_ot(self, val):
        self.ent_ot.delete(0, tk.END)
        self.ent_ot.insert(0, val)

    def bind_name(self, seq, func): self.cb_name.bind(seq, func)
    def bind_time(self, seq, func): 
        self.ent_worktime.bind(seq, func)
        # Shift selection should also trigger any auto-save bindings
        self.cb_shift.bind('<<ComboboxSelected>>', func, add='+')
        if 'FocusOut' in seq or 'Return' in seq:
            # Also trigger on selection from dropdown
            self.ent_worktime.bind('<<ComboboxSelected>>', func, add='+')
            
    def bind_ot(self, seq, func): self.ent_ot.bind(seq, func)

    def update_time_list(self, new_list):
        """Refresh the combobox values with a new list"""
        if hasattr(self, 'ent_worktime'):
            self.ent_worktime['values'] = new_list

class VehicleInspectionWidget(ttk.Frame):
    """
    Dedicated widget for vehicle inspection records with scrollable content
    """
    def __init__(self, parent, theme_bg='#f0f0f0', vehicle_list=None, **kwargs):
        super().__init__(parent, padding=0)
        
        # Create Canvas and Scrollbar for internal scrolling
        # [REFINEMENT] Increased width and height to accommodate the 2x4 table
        self.canvas = tk.Canvas(self, highlightthickness=0, bg=theme_bg, width=680, height=350)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)


        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        # Scroll binding is handled globally in MaterialManager.__init__

        def _on_canvas_configure(event):
            # Update width of inner frame to match canvas
            self.canvas.itemconfig(self.canvas_window, width=event.width)
        self.canvas.bind("<Configure>", _on_canvas_configure)

        # 1. Inspection Items (2x4 Table based on user request)
        chk_frame = ttk.LabelFrame(self.scrollable_frame, text="차량관리 및 안전관리 점검")
        chk_frame.pack(fill='x', padx=5, pady=2)
        
        # Headers (Table Column Labels)
        headers = ["구분", "차량 외부상태", "차량내부 청결상태", "차량내부 청소여부", "이동함 시건장치"]
        for i, h in enumerate(headers):
            colspan = 1 if i == 0 else 2
            col_idx = 0 if i == 0 else 1 + (i-1)*2
            tk.Label(chk_frame, text=h, font=('Malgun Gothic', 8, 'bold'), background=theme_bg).grid(row=0, column=col_idx, columnspan=colspan, padx=2, pady=2)

        self.vars = {}
        # Define Categories and Options
        rows = [("출차시", "out"), ("입차시", "in")]
        categories = [
            ("exterior", ["양호", "불량"]),
            ("cleanliness", ["양호", "불량"]),
            ("cleaning", ["함", "안함"]), # [RESTORED] Changed back to 함/안함 as requested
            ("locking", ["잠금", "안함"])   # [FIX] Changed to 잠금 to match template
        ]

        for r_idx, (r_label, r_key) in enumerate(rows, 1):
            tk.Label(chk_frame, text=r_label, font=('Malgun Gothic', 8), background=theme_bg).grid(row=r_idx, column=0, padx=5, pady=2)
            for c_idx, (c_key, options) in enumerate(categories):
                var = tk.StringVar(value="")
                self.vars[f"{r_key}_{c_key}"] = var
                
                start_col = 1 + (c_idx * 2)
                for o_idx, opt in enumerate(options):
                    # Custom Checkbutton logic to simulate radio behavior (one per category/row)
                    cb = ttk.Checkbutton(chk_frame, text=opt, variable=var, 
                                       onvalue=opt, offvalue="", 
                                       command=lambda v=var, o=opt: self._ensure_exclusive(v, o))
                    cb.grid(row=r_idx, column=start_col + o_idx, padx=2, pady=2, sticky='w')

        # 2. Input Fields
        input_frame = ttk.Frame(self.scrollable_frame)
        input_frame.pack(fill='x', padx=5, pady=2)
        
        ttk.Label(input_frame, text="차량정보:").grid(row=0, column=0, padx=2, pady=2, sticky='e')
        self.cb_vehicle_info = ttk.Combobox(input_frame, width=15)
        self.cb_vehicle_info.grid(row=0, column=1, padx=2, pady=2, sticky='w')
        if vehicle_list is not None:
            self.update_vehicle_list(vehicle_list)
            
        # Focus transitions
        def move_to_mileage(e): self.ent_mileage.focus_set()
        self.cb_vehicle_info.bind('<<ComboboxSelected>>', move_to_mileage)
        self.cb_vehicle_info.bind('<Return>', move_to_mileage)

        ttk.Label(input_frame, text="주행거리 (km):").grid(row=1, column=0, padx=2, pady=2, sticky='e')
        self.ent_mileage = ttk.Entry(input_frame, width=15)
        self.ent_mileage.grid(row=1, column=1, padx=2, pady=2, sticky='w')
        
        def on_mileage_return(e):
            self.format_mileage()
            self.ent_remarks.focus_set()
            
        self.ent_mileage.bind('<Return>', on_mileage_return)
        self.ent_mileage.bind('<FocusOut>', lambda e: self.format_mileage())

        ttk.Label(input_frame, text="비고:").grid(row=2, column=0, padx=2, pady=2, sticky='e')
        self.ent_remarks = ttk.Entry(input_frame)
        self.ent_remarks.grid(row=2, column=1, padx=2, pady=2, sticky='ew')
        input_frame.grid_columnconfigure(1, weight=1)

        # [NEW] Add Standalone Save Button
        self.on_save_callback = kwargs.get('on_save')
        if self.on_save_callback:
            btn_save = ttk.Button(self.scrollable_frame, text="🚗 차량 정보만 개별 저장", 
                                 command=self.trigger_save, style='Accent.TButton' if 'Accent.TButton' in getattr(parent, 'style_names', []) else 'TButton')
            btn_save.pack(fill='x', padx=5, pady=5)
            
        # 3. Photo Section [NEW]
        photo_frame = ttk.LabelFrame(self.scrollable_frame, text="차량 사진 (클릭 시 크게 보기)")
        photo_frame.pack(fill='x', padx=5, pady=2)
        
        # [LARGER] Increased preview area size
        self.photo_canvas = tk.Label(photo_frame, text="사진 없음", background="#e0e0e0", width=60, height=15, cursor='hand2')
        self.photo_canvas.pack(pady=5)
        self.photo_canvas.bind("<Button-1>", lambda e: self.open_full_photo())
        
        self.photo_path = ""
        self.photo_image = None # Keep reference
        
        btn_photo_f = ttk.Frame(photo_frame)
        btn_photo_f.pack(fill='x', pady=2)
        
        ttk.Button(btn_photo_f, text="📸 사진 추가/변경", command=self.add_photo).pack(side='left', expand=True, fill='x', padx=2)
        ttk.Button(btn_photo_f, text="🗑️ 사진 삭제", command=self.clear_photo).pack(side='left', expand=True, fill='x', padx=2)

    def _ensure_exclusive(self, var, current_val):
        """Helper to ensure only one value is selected if needed (though StringVar handles it naturally)"""
        # This is primarily to handle clicking an already selected check to uncheck it if desired,
        # but ttk.Checkbutton with variable already does this for on/off values.
        pass

    def trigger_save(self):
        """Invoke the save callback provided by MaterialManager"""
        if self.on_save_callback:
            self.on_save_callback(self)

    def update_vehicle_list(self, new_list):
        """Update the combobox values (native dropdown)"""
        self.cb_vehicle_info['values'] = new_list

    def format_mileage(self, event=None):
        """Autoformat mileage with commas"""
        try:
            val = self.ent_mileage.get().strip().replace(',', '')
            if not val: return
            # Handle possible float if someone type 123.4
            if '.' in val:
                num = float(val)
                formatted = f"{num:,.1f}"
                if formatted.endswith('.0'): formatted = formatted[:-2]
            else:
                num = int(val)
                formatted = f"{num:,}"
            
            self.ent_mileage.delete(0, tk.END)
            self.ent_mileage.insert(0, formatted)
        except:
            pass

    def get_data(self):
        # Collect all inspection variables
        data = {key: var.get() for key, var in self.vars.items()}
        data['vehicle_info'] = self.cb_vehicle_info.get().strip()
        data['mileage'] = self.ent_mileage.get().strip()
        data['remarks'] = self.ent_remarks.get().strip()
        
        data['_raw_mileage'] = data['mileage'].replace(',', '')
        data['photo_path'] = self.photo_path
        return data

    def add_photo(self):
        """Open file dialog and copy photo to storage"""
        file_path = filedialog.askopenfilename(
            title="차량 사진 선택",
            filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp *.gif")]
        )
        if not file_path: return
        
        try:
            # Create directory if missing
            photo_dir = r"c:\Users\-\OneDrive\바탕 화면\home\data\vehicle_photos"
            if not os.path.exists(photo_dir):
                os.makedirs(photo_dir)
                
            v_no = self.cb_vehicle_info.get().strip()
            if not v_no:
                messagebox.showwarning("경고", "사진을 저장하기 전에 차량정보(차량번호)를 먼저 입력해주세요.")
                return
                
            # Clean filename
            clean_v_no = "".join(c for c in v_no if c.isalnum() or c in (' ', '-', '_')).strip()
            timestamp = int(time.time())
            ext = os.path.splitext(file_path)[1]
            new_filename = f"{clean_v_no}_{timestamp}{ext}"
            new_path = os.path.join(photo_dir, new_filename)
            
            # Copy file
            import shutil
            shutil.copy2(file_path, new_path)
            
            # Update and display
            self.photo_path = new_path
            self.display_photo()
            
        except Exception as e:
            messagebox.showerror("오류", f"사진을 저장하는 중 오류가 발생했습니다: {e}")

    def display_photo(self):
        """Load and display the photo in the widget"""
        if not self.photo_path or not os.path.exists(self.photo_path):
            self.photo_canvas.config(image='', text="사진 없음")
            return
            
        try:
            img = Image.open(self.photo_path)
            # [LARGER] Increased thumbnail size for better visibility
            img.thumbnail((600, 400))
            self.photo_image = ImageTk.PhotoImage(img)
            self.photo_canvas.config(image=self.photo_image, text="")
        except Exception as e:
            print(f"DEBUG: Error displaying photo: {e}")
            self.photo_canvas.config(image='', text="이미지 로드 오류")

    def open_full_photo(self):
        """Open the photo in a separate large window"""
        if not self.photo_path or not os.path.exists(self.photo_path): return
        
        full_win = tk.Toplevel(self)
        full_win.title("차량 사진 크게 보기")
        full_win.geometry("1000x800")
        
        try:
            img = Image.open(self.photo_path)
            # Maximize for the window
            img.thumbnail((1200, 900))
            full_photo = ImageTk.PhotoImage(img)
            
            # Keep reference to avoid GC
            full_win.full_photo = full_photo 
            
            canvas = tk.Canvas(full_win, bg='black')
            canvas.pack(fill='both', expand=True)
            
            canvas.create_image(500, 400, image=full_photo, anchor='center')
            
            ttk.Button(full_win, text="닫기", command=full_win.destroy).pack(pady=5)
        except Exception as e:
            messagebox.showerror("오류", f"이미지를 여는 중 오류가 발생했습니다: {e}")

    def clear_photo(self):
        """Remove photo from widget (doesn't delete file)"""
        self.photo_path = ""
        self.photo_image = None
        self.photo_canvas.config(image='', text="사진 없음")

    def reset_fields(self):
        """Clear all input fields for the next vehicle entry and set focus"""
        try:
            self.cb_vehicle_info.set('')
            self.ent_mileage.delete(0, tk.END)
            self.ent_remarks.delete(0, tk.END)
            for var in self.vars.values():
                var.set("") # Clear StringVar
            self.cb_vehicle_info.focus_set()
        except:
            pass
        
    def set_data(self, data):
        if not data: return
        for key, val in data.items():
            if key in self.vars:
                self.vars[key].set(val)
            elif key == 'vehicle_info':
                self.cb_vehicle_info.set(val)
            elif key == 'mileage':
                self.ent_mileage.delete(0, tk.END)
                self.ent_mileage.insert(0, val)
            elif key == 'remarks':
                self.ent_remarks.delete(0, tk.END)
                self.ent_remarks.insert(0, val)
            elif key == 'photo_path':
                self.photo_path = val
                self.display_photo()


def enable_column_resize(frame, num_cols, header_row=0, edge_px=6):
    """
    frame 내부 grid 테이블의 헤더 셀(header_row) 오른쪽 경계를 드래그해
    해당 컬럼과 오른쪽 인접 컬럼의 너비를 조절하는 기능을 추가합니다.

    - frame  : ttk.Frame, grid 레이아웃으로 컬럼이 구성된 프레임
    - num_cols : 컬럼 수
    - header_row : 헤더가 위치한 row 번호 (기본 0)
    - edge_px : 경계 감지 영역 (픽셀)
    """
    _drag = {'col': None, 'start_x': 0, 'start_w': 0, 'next_w': 0}

    def get_col_width(col):
        """현재 컬럼의 실제 픽셀 너비를 반환"""
        try:
            info = frame.grid_columnconfigure(col)
            minsz = info.get('minsize', 0)
            if minsz and minsz > 0:
                return minsz
            for widget in frame.grid_slaves(row=header_row, column=col):
                w = widget.winfo_width()
                if w > 1:
                    return w
            return 80
        except:
            return 80

    def on_motion(event):
        """드래그 중이 아닐 때만 커서 모양 변경"""
        if _drag['col'] is not None:
            return  # 이미 드래그 중이면 무시
        widget = event.widget
        x = event.x
        w = widget.winfo_width()
        if w - edge_px <= x <= w:
            widget.configure(cursor='sb_h_double_arrow')
        else:
            widget.configure(cursor='')

    def on_leave(event):
        if _drag['col'] is None:
            event.widget.configure(cursor='')

    def on_press(event):
        widget = event.widget
        x = event.x
        w = widget.winfo_width()
        if w - edge_px <= x <= w:
            info = widget.grid_info()
            col = info.get('column', -1)
            if col < 0 or col >= num_cols - 1:
                return
            _drag['col'] = col
            _drag['widget'] = widget
            _drag['start_x'] = event.x_root
            _drag['start_w'] = get_col_width(col)
            _drag['next_w'] = get_col_width(col + 1)
            # grab_set 제거 - 스크롤 방해 방지
        else:
            _drag['col'] = None

    def on_drag(event):
        if _drag['col'] is None:
            return
        col = _drag['col']
        dx = event.x_root - _drag['start_x']
        new_w = max(30, _drag['start_w'] + dx)
        new_next = max(30, _drag['next_w'] - dx)
        frame.grid_columnconfigure(col, minsize=new_w, weight=0)
        frame.grid_columnconfigure(col + 1, minsize=new_next, weight=0)

    def on_release(event):
        _drag['col'] = None
        # 커서 원상복구
        try:
            w = event.widget
            w.configure(cursor='')
        except:
            pass

    # 헤더 row(0) 의 모든 위젯에 바인딩
    def bind_headers():
        for col in range(num_cols):
            for widget in frame.grid_slaves(row=header_row, column=col):
                widget.bind('<Motion>', on_motion, add='+')
                widget.bind('<Leave>', on_leave, add='+')
                widget.bind('<ButtonPress-1>', on_press, add='+')
                widget.bind('<B1-Motion>', on_drag, add='+')
                widget.bind('<ButtonRelease-1>', on_release, add='+')

    # 위젯이 아직 렌더되기 전일 수 있으므로 idle 후 바인딩
    frame.after_idle(bind_headers)


class LaborCostDetailWidget(ttk.Frame):
    """
    Detailed labor cost calculation widget with three sections: 
    1) Regular Work (정시근무), 2) Special Work (특별근무), and 3) Base Salary Reference (기준급여)
    """
    def __init__(self, parent, on_change_callback=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.on_change_callback = on_change_callback
        
        # Rankings for Table 1
        self.ranks = ["이사", "부장", "차장", "과장", "대리", "계장", "주임", "기사"]
        # Shift types for Table 2
        self.special_types = ["연장근무", "야간근무", "휴일근무"]
        
        # Base Salaries for Reference (Table 3) - Dynamically loaded from master
        self.base_salaries = {}
        
        # Resolve MaterialManager to get rates
        master = parent
        while master and not hasattr(master, 'get_base_salaries'):
            master = getattr(master, 'master', None)
        
        if master:
            self.base_salaries = master.get_base_salaries()
        else:
             self.base_salaries = {
                "이사": 55250000, "부장": 55250000, "차장": 47670000, "과장": 41170000,
                "대리": 37920000, "계장": 34670000, "주임": 31420000, "기사": 29250000
            }
        
        self.entries = {} # Key -> Rank/Type -> Column -> Entry
        self.totals = {}  # Key -> Rank/Type -> Label
        
        self._create_widgets()

    def get_total_cost(self):
        """[FINAL_FIX] Robustly get total cost from label at class level"""
        try:
            raw_text = getattr(self, 'lbl_grand_total', None)
            if raw_text:
                txt = raw_text.cget('text')
                val = "".join(c for c in txt if c.isdigit() or c == '.')
                return float(val or 0)
            return 0.0
        except:
            return 0.0

    def _create_widgets(self):
        style = ttk.Style()
        style.configure("LaborHeader.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#e0e0e0', relief='solid')
        style.configure("LaborTotal.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#fff9c4', relief='solid')
        style.configure("RefTitle.TLabel", font=('Malgun Gothic', 10, 'bold'), foreground='red', background='#fce4ec', relief='solid')
        
        # Main Container
        main_container = ttk.Frame(self)
        main_container.pack(fill='both', expand=True)
        
        # Left Side (Calculation)
        calc_frame = ttk.Frame(main_container)
        calc_frame.pack(side='left', fill='both', expand=True)
        
        # Right Side (Reference Table)
        ref_frame = ttk.Frame(main_container, padding=(20, 40, 0, 0))
        ref_frame.pack(side='right', fill='y')

        # --- Section 1: Regular Work (Left) ---
        ttk.Label(calc_frame, text="1) 정시근무 (240일/년)", font=('Malgun Gothic', 11, 'bold')).pack(anchor='w', pady=(10, 5))
        
        table1_frame = ttk.Frame(calc_frame)
        table1_frame.pack(fill='x')
        
        headers1 = ["구분", "직급", "투입인원(명)", "투입일수/인(일)", "단가/일", "사전원가가액"]
        for j, h in enumerate(headers1):
            lbl = ttk.Label(table1_frame, text=h, style="LaborHeader.TLabel", padding=5, anchor='center')
            lbl.grid(row=0, column=j, sticky='nsew')
            table1_frame.grid_columnconfigure(j, weight=1)
        enable_column_resize(table1_frame, len(headers1))

        merge_lbl = ttk.Label(table1_frame, text="정시근무\n(240일/년)", relief='solid', anchor='center', padding=10)
        merge_lbl.grid(row=1, column=0, rowspan=len(self.ranks), sticky='nsew')

        for i, rank in enumerate(self.ranks):
            row = i + 1
            ttk.Label(table1_frame, text=rank, relief='solid', anchor='center', padding=5).grid(row=row, column=1, sticky='nsew')
            
            self.entries[rank] = {}
            ent_personnel = ttk.Entry(table1_frame, width=10, justify='center')
            ent_personnel.grid(row=row, column=2, sticky='nsew')
            ent_personnel.bind("<KeyRelease>", lambda e, r=rank: self._on_input_change(r))
            self.entries[rank]['personnel'] = ent_personnel
            
            ent_days = ttk.Entry(table1_frame, width=10, justify='center')
            ent_days.grid(row=row, column=3, sticky='nsew')
            ent_days.bind("<KeyRelease>", lambda e, r=rank: self._on_input_change(r))
            self.entries[rank]['period'] = ent_days
            
            ent_price = ttk.Entry(table1_frame, width=15, justify='right')
            ent_price.grid(row=row, column=4, sticky='nsew')
            ent_price.bind("<KeyRelease>", lambda e, r=rank: self._on_input_change(r))
            self.entries[rank]['unit_price'] = ent_price
            
            # [NEW] Default value from base salary / 240
            daily_rate = round(self.base_salaries.get(rank, 0) / 240)
            if daily_rate > 0:
                ent_price.insert(0, f"{daily_rate:,.0f}")
            
            lbl_subtotal = ttk.Label(table1_frame, text="0", relief='solid', anchor='e', padding=5)
            lbl_subtotal.grid(row=row, column=5, sticky='nsew')
            self.totals[rank] = lbl_subtotal

        # Table 1 Subtotal Row
        row_t1_sum = len(self.ranks) + 1
        ttk.Label(table1_frame, text="소계", style="LaborHeader.TLabel", anchor='center', padding=5).grid(row=row_t1_sum, column=0, columnspan=2, sticky='nsew')
        self.lbl_t1_personnel_sum = ttk.Label(table1_frame, text="0 명", style="LaborHeader.TLabel", anchor='center')
        self.lbl_t1_personnel_sum.grid(row=row_t1_sum, column=2, sticky='nsew')
        self.lbl_t1_days_sum = ttk.Label(table1_frame, text="0 일", style="LaborHeader.TLabel", anchor='center')
        self.lbl_t1_days_sum.grid(row=row_t1_sum, column=3, sticky='nsew')
        ttk.Label(table1_frame, text="", style="LaborHeader.TLabel").grid(row=row_t1_sum, column=4, sticky='nsew')
        self.lbl_t1_cost_sum = ttk.Label(table1_frame, text="0", style="LaborHeader.TLabel", anchor='e', padding=5)
        self.lbl_t1_cost_sum.grid(row=row_t1_sum, column=5, sticky='nsew')

        # --- Section 2: Special Work (Left) ---
        ttk.Label(calc_frame, text="2) 특별근무", font=('Malgun Gothic', 11, 'bold')).pack(anchor='w', pady=(20, 5))
        
        table2_frame = ttk.Frame(calc_frame)
        table2_frame.pack(fill='x')
        
        headers2 = ["구분", "형태", "투입인원(명)", "투입시간/인(시간)", "단가", "사전원가가액"]
        for j, h in enumerate(headers2):
            lbl = ttk.Label(table2_frame, text=h, style="LaborHeader.TLabel", padding=5, anchor='center')
            lbl.grid(row=0, column=j, sticky='nsew')
            table2_frame.grid_columnconfigure(j, weight=1)
        enable_column_resize(table2_frame, len(headers2))

        merge_lbl2 = ttk.Label(table2_frame, text="특별근무", relief='solid', anchor='center', padding=10)
        merge_lbl2.grid(row=1, column=0, rowspan=len(self.special_types), sticky='nsew')

        for i, stype in enumerate(self.special_types):
            row = i + 1
            ttk.Label(table2_frame, text=stype, relief='solid', anchor='center', padding=5).grid(row=row, column=1, sticky='nsew')
            
            self.entries[stype] = {}
            ent_personnel = ttk.Entry(table2_frame, width=10, justify='center')
            ent_personnel.grid(row=row, column=2, sticky='nsew')
            ent_personnel.bind("<KeyRelease>", lambda e, s=stype: self._on_input_change(s))
            self.entries[stype]['personnel'] = ent_personnel
            
            ent_hours = ttk.Entry(table2_frame, width=10, justify='center')
            ent_hours.grid(row=row, column=3, sticky='nsew')
            ent_hours.bind("<KeyRelease>", lambda e, s=stype: self._on_input_change(s))
            self.entries[stype]['period'] = ent_hours
            
            ent_price = ttk.Entry(table2_frame, width=15, justify='right')
            ent_price.grid(row=row, column=4, sticky='nsew')
            ent_price.bind("<KeyRelease>", lambda e, s=stype: self._on_input_change(s))
            self.entries[stype]['unit_price'] = ent_price
            
            # [NEW] Default OT Rates
            ot_rates = {"연장근무": 4000, "야간근무": 5000, "휴일근무": 7500}
            if stype in ot_rates:
                ent_price.insert(0, f"{ot_rates[stype]:,.0f}")
            
            lbl_subtotal = ttk.Label(table2_frame, text="0", relief='solid', anchor='e', padding=5)
            lbl_subtotal.grid(row=row, column=5, sticky='nsew')
            self.totals[stype] = lbl_subtotal

        # Table 2 Subtotal Row
        row_t2_sum = len(self.special_types) + 1
        ttk.Label(table2_frame, text="소계", style="LaborHeader.TLabel", anchor='center', padding=5).grid(row=row_t2_sum, column=0, columnspan=2, sticky='nsew')
        self.lbl_t2_personnel_sum = ttk.Label(table2_frame, text="0 명", style="LaborHeader.TLabel", anchor='center')
        self.lbl_t2_personnel_sum.grid(row=row_t2_sum, column=2, sticky='nsew')
        self.lbl_t2_hours_sum = ttk.Label(table2_frame, text="0 시간", style="LaborHeader.TLabel", anchor='center')
        self.lbl_t2_hours_sum.grid(row=row_t2_sum, column=3, sticky='nsew')
        ttk.Label(table2_frame, text="", style="LaborHeader.TLabel").grid(row=row_t2_sum, column=4, sticky='nsew')
        self.lbl_t2_cost_sum = ttk.Label(table2_frame, text="0", style="LaborHeader.TLabel", anchor='e', padding=5)
        self.lbl_t2_cost_sum.grid(row=row_t2_sum, column=5, sticky='nsew')

        # --- Section 3: Base Salary Reference (Right) ---
        ttk.Label(ref_frame, text="고정금액(변경불가)", style="RefTitle.TLabel", padding=5).pack(fill='x')
        
        ref_table = ttk.Frame(ref_frame)
        ref_table.pack(fill='x')
        
        ttk.Label(ref_table, text="직급", style="LaborHeader.TLabel", padding=5, width=10, anchor='center').grid(row=0, column=0, sticky='nsew')
        ttk.Label(ref_table, text="기준급여", style="LaborHeader.TLabel", padding=5, width=15, anchor='center').grid(row=0, column=1, sticky='nsew')
        
        # Display ranks (skipping 이사 as it's not in the reference list image, but we have the data)
        display_ranks = ["부장", "차장", "과장", "대리", "계장", "주임", "기사"]
        for i, rank in enumerate(display_ranks):
            row = i + 1
            ttk.Label(ref_table, text=rank, relief='solid', padding=5, anchor='center').grid(row=row, column=0, sticky='nsew')
            salary = self.base_salaries.get(rank, 0)
            ttk.Label(ref_table, text=f"{salary:,.0f}", relief='solid', padding=5, anchor='e').grid(row=row, column=1, sticky='nsew')

        ttk.Button(ref_frame, text="기준급여 일괄 적용", command=self.apply_base_salaries).pack(pady=10, fill='x')

        # --- Grand Total ---
        total_frame = ttk.Frame(calc_frame)
        total_frame.pack(fill='x', pady=(10, 0))
        ttk.Label(total_frame, text="인건비 합계", style="LaborTotal.TLabel", anchor='center', padding=10).pack(side='left', fill='x', expand=True)
        self.lbl_grand_total = ttk.Label(total_frame, text="0", style="LaborTotal.TLabel", anchor='e', font=('Malgun Gothic', 12, 'bold'), padding=10)
        self.lbl_grand_total.pack(side='right', fill='x', expand=True)

    def apply_base_salaries(self):
        """Reset Daily Unit Price to the standard (Base Salary / 240)"""
        for rank, salary in self.base_salaries.items():
            if rank in self.entries:
                daily_rate = round(salary / 240)
                self.entries[rank]['unit_price'].delete(0, tk.END)
                self.entries[rank]['unit_price'].insert(0, f"{daily_rate:,.0f}")
                self._on_input_change(rank)
        messagebox.showinfo("적용 완료", "기준급여에 따른 일일 단가가 적용되었습니다.")

    def get_total_cost(self):
        """Retrieve the final calculated labor cost as a float"""
        try:
            val = self.lbl_grand_total.cget('text').replace('₩', '').replace(',', '').replace(' ', '').strip()
            return float(val or 0)
        except:
            return 0.0

    def _to_f(self, val):
        try:
            return float(str(val).replace(',', '') or 0)
        except:
            return 0.0

    def _on_input_change(self, key):
        # Calculate row total
        personnel = self._to_f(self.entries[key]['personnel'].get())
        period = self._to_f(self.entries[key]['period'].get())
        price = self._to_f(self.entries[key]['unit_price'].get())
        
        row_total = personnel * period * price
        self.totals[key].config(text=f"{row_total:,.0f}")
        
        self.calculate_all()

    def calculate_all(self):
        # Table 1 Totals
        t1_personnel = 0
        t1_days = 0
        t1_cost = 0
        for rank in self.ranks:
            p = self._to_f(self.entries[rank]['personnel'].get())
            d = self._to_f(self.entries[rank]['period'].get())
            c = self._to_f(self.totals[rank].cget('text'))
            t1_personnel += p
            t1_days += d
            t1_cost += c
        
        self.lbl_t1_personnel_sum.config(text=f"{t1_personnel:g} 명")
        self.lbl_t1_days_sum.config(text=f"{t1_days:g} 일")
        self.lbl_t1_cost_sum.config(text=f"{t1_cost:,.0f}")

        # Table 2 Totals
        t2_personnel = 0
        t2_hours = 0
        t2_cost = 0
        for stype in self.special_types:
            p = self._to_f(self.entries[stype]['personnel'].get())
            h = self._to_f(self.entries[stype]['period'].get())
            c = self._to_f(self.totals[stype].cget('text'))
            t2_personnel += p
            t2_hours += h
            t2_cost += c
            
        self.lbl_t2_personnel_sum.config(text=f"{t2_personnel:g} 명")
        self.lbl_t2_hours_sum.config(text=f"{t2_hours:g} 시간")
        self.lbl_t2_cost_sum.config(text=f"{t2_cost:,.0f}")

        grand_total = t1_cost + t2_cost
        self.lbl_grand_total.config(text=f"₩ {grand_total:,.0f}")
        
        if self.on_change_callback:
            self.on_change_callback(grand_total)

    def get_data(self):
        """Export all entry values as a dictionary"""
        data = {}
        for key, widgets in self.entries.items():
            data[key] = {
                'personnel': widgets['personnel'].get(),
                'period': widgets['period'].get(),
                'unit_price': widgets['unit_price'].get()
            }
        return data

    def set_data(self, data):
        """Populate entries from a dictionary"""
        if not data or not isinstance(data, dict):
            self.reset()
            return
            
        for key, values in data.items():
            if key in self.entries:
                self.entries[key]['personnel'].delete(0, tk.END); self.entries[key]['personnel'].insert(0, values.get('personnel', ''))
                self.entries[key]['period'].delete(0, tk.END); self.entries[key]['period'].insert(0, values.get('period', ''))
                self.entries[key]['unit_price'].delete(0, tk.END); self.entries[key]['unit_price'].insert(0, values.get('unit_price', ''))
        
        # Trigger all row calculations
        for key in list(self.ranks) + list(self.special_types):
            self._on_input_change(key)

    def reset(self):
        """Clear all entries"""
        for key, widgets in self.entries.items():
            widgets['personnel'].delete(0, tk.END)
            widgets['period'].delete(0, tk.END)
            widgets['unit_price'].delete(0, tk.END)
        self.calculate_all()

class MaterialCostDetailWidget(ttk.Frame):
    """
    Detailed material cost calculation widget.
    Columns: Item (품목), Spec (사양), Quantity (수량), Unit (규격), Price (단가), Amount (사전원가가액)
    """
    def __init__(self, parent, on_change_callback=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.on_change_callback = on_change_callback
        
        # Default Items from settings_df or fallback
        self.default_items = []
        
        # Resolve MaterialManager to get rates
        master = parent
        while master and not hasattr(master, 'get_material_defaults'):
            master = getattr(master, 'master', None)
        
        if master:
            self.default_items = master.get_material_defaults()
        else:
            self.default_items = [
                ("PT 약품", "세척제", "CAN", 1500), ("PT 약품", "침투제", "CAN", 2300),
                ("PT 약품", "현상제", "CAN", 2000), ("MT 약품", "백색페인트", "CAN", 2350),
                ("MT 약품", "흑색자분", "CAN", 1800), ("방사선투과검사 필름", "MX125", "매", 990),
                ("글리세린", "20L", "통", 100000), ("필름 현상액", "3L", "통", 16500),
                ("필름 정착액", "3L", "통", 16500), ("수적방지액", "200mL", "통", 2500)
            ]
        
        self.entries = [] # List of dicts for each row: {'item': lbl, 'spec': lbl, 'qty': ent, 'unit': lbl, 'price': ent, 'amount': lbl}
        self._create_widgets()

    def _create_widgets(self):
        style = ttk.Style()
        style.configure("MatHeader.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#e0e0e0', relief='solid')
        style.configure("MatTotal.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#ffff00', relief='solid') # Yellow as in image
        style.configure("MatRef.TLabel", font=('Malgun Gothic', 9), background='#f5f5f5', relief='solid')

        # Main Container
        main_container = ttk.Frame(self)
        main_container.pack(fill='both', expand=True)

        # Left Side (Calc)
        calc_frame = ttk.Frame(main_container)
        calc_frame.pack(side='left', fill='both', expand=True)

        # Right Side (Ref)
        ref_frame = ttk.Frame(main_container, padding=(20, 40, 0, 0))
        ref_frame.pack(side='right', fill='y')

        # --- Section 1: Tables (Left) ---
        ttk.Label(calc_frame, text="2) 재료비", font=('Malgun Gothic', 11, 'bold')).pack(anchor='w', pady=(10, 5))
        
        table_frame = ttk.Frame(calc_frame)
        table_frame.pack(fill='x')
        
        headers = ["품목", "사양", "수량", "규격", "단가", "사전원가가액"]
        widths = [20, 15, 10, 8, 15, 20]
        for j, (h, w) in enumerate(zip(headers, widths)):
            lbl = ttk.Label(table_frame, text=h, style="MatHeader.TLabel", padding=5, anchor='center', width=w)
            lbl.grid(row=0, column=j, sticky='nsew')
            table_frame.grid_columnconfigure(j, weight=1 if j in [0, 5] else 0)
        enable_column_resize(table_frame, len(headers))

        for i, (item, spec, unit, price) in enumerate(self.default_items):
            row = i + 1
            # Item
            ttk.Label(table_frame, text=item, relief='solid', padding=5, anchor='center').grid(row=row, column=0, sticky='nsew')
            # Spec
            ttk.Label(table_frame, text=spec, relief='solid', padding=5, anchor='center').grid(row=row, column=1, sticky='nsew')
            
            row_widgets = {}
            # Quantity
            ent_qty = ttk.Entry(table_frame, width=10, justify='center')
            ent_qty.grid(row=row, column=2, sticky='nsew')
            ent_qty.bind("<KeyRelease>", lambda e, idx=i: self._on_input_change(idx))
            row_widgets['qty'] = ent_qty
            
            # Unit
            ttk.Label(table_frame, text=unit, relief='solid', padding=5, anchor='center').grid(row=row, column=3, sticky='nsew')
            
            # Unit Price
            ent_price = ttk.Entry(table_frame, width=15, justify='right')
            ent_price.grid(row=row, column=4, sticky='nsew')
            ent_price.insert(0, f"{price:,.0f}")
            ent_price.bind("<KeyRelease>", lambda e, idx=i: self._on_input_change(idx))
            row_widgets['price'] = ent_price
            
            # Amount
            lbl_amount = ttk.Label(table_frame, text="0", relief='solid', anchor='e', padding=5)
            lbl_amount.grid(row=row, column=5, sticky='nsew')
            row_widgets['amount'] = lbl_amount
            
            self.entries.append(row_widgets)

        # Footer 1: Total
        row_footer = len(self.default_items) + 1
        ttk.Label(table_frame, text="재료비 합계", style="MatTotal.TLabel", anchor='center', padding=5).grid(row=row_footer, column=0, columnspan=5, sticky='nsew')
        self.lbl_mat_total = ttk.Label(table_frame, text="₩ 0", style="MatTotal.TLabel", anchor='e', padding=5)
        self.lbl_mat_total.grid(row=row_footer, column=5, sticky='nsew')

        # Footer 2: VAT
        row_vat = len(self.default_items) + 2
        ttk.Label(table_frame, text="부가세", relief='solid', anchor='center', padding=5).grid(row=row_vat, column=4, sticky='nsew')
        self.lbl_mat_vat = ttk.Label(table_frame, text="0", relief='solid', anchor='e', padding=5)
        self.lbl_mat_vat.grid(row=row_vat, column=5, sticky='nsew')

        # --- Section 2: Ref Info (Right) ---
        ref_table = ttk.Frame(ref_frame)
        ref_table.pack(fill='x')

        # Headers for reference
        ref_data = [
            ("지에스켐", "30M"),
            ("지에스켐", ""),
            ("지에스켐", ""),
            ("지에스켐", "32M"),
            ("지에스켐", ""),
            ("한스", "3 1/3 x 12\" (1850 매)"),
            ("한스", "3 1/3 x 6\" (313 매)"),
            ("한스", "총 매수: 2,163"),
            ("경도", "500매 기준 / 1회 4통 / 5회 교환"),
            ("나우", "500매 기준 / 1회 2동 / 5회 교환")
        ]

        for i, (title, val) in enumerate(ref_data):
            row = i
            ttk.Label(ref_table, text=title, width=10, anchor='center', style="MatRef.TLabel", padding=3).grid(row=row, column=0, sticky='nsew')
            ttk.Label(ref_table, text=val, width=25, anchor='w', style="MatRef.TLabel", padding=3).grid(row=row, column=1, sticky='nsew')

    def _on_input_change(self, idx):
        widgets = self.entries[idx]
        qty = self._to_f(widgets['qty'].get())
        price = self._to_f(widgets['price'].get())
        
        amount = qty * price
        widgets['amount'].config(text=f"{amount:,.0f}")
        
        self.calculate_all()

    def calculate_all(self):
        total_mat = 0.0
        for widgets in self.entries:
            qty = self._to_f(widgets['qty'].get())
            price = self._to_f(widgets['price'].get())
            amount = qty * price
            widgets['amount'].config(text=f"{amount:,.0f}")
            total_mat += amount
            
        vat = total_mat * 0.1
        self.lbl_mat_total.config(text=f"₩ {total_mat:,.0f}")
        self.lbl_mat_vat.config(text=f"{vat:,.0f}")
        
        if self.on_change_callback:
            self.on_change_callback(total_mat)

    def _to_f(self, val):
        try:
            return float(str(val).replace(',', '') or 0)
        except:
            return 0.0

    def get_data(self):
        """Export entry values"""
        data = []
        for widgets in self.entries:
            data.append({
                'qty': widgets['qty'].get(),
                'price': widgets['price'].get()
            })
        return data

    def set_data(self, data):
        """Populate entries"""
        if not data or not isinstance(data, list):
            self.reset()
            return
            
        for i, val in enumerate(data):
            if i < len(self.entries):
                self.entries[i]['qty'].delete(0, tk.END); self.entries[i]['qty'].insert(0, val.get('qty', ''))
                self.entries[i]['price'].delete(0, tk.END); self.entries[i]['price'].insert(0, val.get('price', ''))
        
        self.calculate_all()

    def reset(self):
        """Clear quantities and restore default prices"""
        for i, widgets in enumerate(self.entries):
            widgets['qty'].delete(0, tk.END)
            # Restore default price
            default_price = self.default_items[i][3]
            widgets['price'].delete(0, tk.END); widgets['price'].insert(0, f"{default_price:,.0f}")
            
        self.calculate_all()

class ExpenseProfitDetailWidget(ttk.Frame):
    """
    Comprehensive expense and profit calculation widget.
    Sections: 1) Site Expenses, 2) Rental, 3) Outsource, 4) Insurance, 5) Depreciation, 6) Indirect Cost, 7) Profit
    """
    def __init__(self, parent, on_change_callback=None, get_labor_total_func=None, get_material_total_func=None, get_revenue_func=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.on_change_callback = on_change_callback
        self.get_labor_total = get_labor_total_func
        self.get_material_total = get_material_total_func
        self.get_revenue = get_revenue_func
        
        self.entries = {
            'site_expense': [], # list of dicts
            'rental': [],
            'outsource': [],
            'depreciation': []
        }
        
        # Resolve MaterialManager to get rates
        self.master_app = parent
        while self.master_app and not hasattr(self.master_app, 'get_expense_defaults'):
            self.master_app = getattr(self.master_app, 'master', None)
            
        self._create_widgets()

    def _create_widgets(self):
        style = ttk.Style()
        style.configure("ExpHeader.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#e0e0e0', relief='solid')
        style.configure("ExpTotal.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#ffff00', relief='solid')
        style.configure("Margin.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#90ee90', relief='solid') # Light green for profit

        # --- Section 1: Site Expenses ---
        ttk.Label(self, text="3) 경비", font=('Malgun Gothic', 11, 'bold')).pack(anchor='w', pady=(10, 5))
        
        s1_frame = ttk.LabelFrame(self, text="(1) 현장 경비")
        s1_frame.pack(fill='x', pady=5)
        
        self.s1_table = ttk.Frame(s1_frame)
        self.s1_table.pack(fill='x')
        
        headers = ["구분", "내용", "인원수", "수량", "규격", "단가", "사전원가가액"]
        widths = [15, 20, 8, 8, 8, 15, 20]
        for j, (h, w) in enumerate(zip(headers, widths)):
            ttk.Label(self.s1_table, text=h, style="ExpHeader.TLabel", padding=5, anchor='center', width=w).grid(row=0, column=j, sticky='nsew')
            self.s1_table.grid_columnconfigure(j, weight=1 if j in [1, 6] else 0)
        enable_column_resize(self.s1_table, len(headers))

        # Default Site Expenses - Dynamically loaded from master
        defaults_s1 = []
        if self.master_app:
            defaults_s1 = self.master_app.get_expense_defaults()
        else:
            defaults_s1 = [
                ("차량유지비", "주유, 수리, 통행, 주차 등", "N/A", 1, "일", 5000),
                ("소모품비", "장갑,일회용 작업복외", "N/A", 1, "일", 500),
                ("복리후생비", "생수, 음료 외 기타", "N/A", 1, "일", 1667),
                ("Se-175", "방사성동위원소 구매", "N/A", 1, "EA", 35714)
            ]
        
        for i, (cat, cont, ppl, qty, unit, price) in enumerate(defaults_s1):
            self._add_row_s1(cat, cont, ppl, qty, unit, price)

        # --- Section 2: Rental Costs ---
        s2_frame = ttk.LabelFrame(self, text="(2) 장비/차량 임차료")
        s2_frame.pack(fill='x', pady=5)
        
        self.s2_table = ttk.Frame(s2_frame)
        self.s2_table.pack(fill='x')
        
        headers2 = ["구분", "사양", "수량", "사용기간", "기간단위", "단가/월,대", "사전원가가액"]
        for j, h in enumerate(headers2):
            ttk.Label(self.s2_table, text=h, style="ExpHeader.TLabel", padding=5, anchor='center', width=widths[j]).grid(row=0, column=j, sticky='nsew')
            self.s2_table.grid_columnconfigure(j, weight=1 if j in [1, 6] else 0)
        enable_column_resize(self.s2_table, len(headers2))
            
        # Add 3 empty rows by default
        for _ in range(3):
            self._add_row_s2()

        # --- Section 3: Outsource Costs ---
        s3_frame = ttk.LabelFrame(self, text="(3) 외주비/잡급")
        s3_frame.pack(fill='x', pady=5)
        
        self.s3_table = ttk.Frame(s3_frame)
        self.s3_table.pack(fill='x')
        
        headers3 = ["구분", "작업내용", "공수", "단가", "사전원가가액"]
        widths3 = [15, 30, 10, 15, 20]
        for j, h in enumerate(headers3):
            ttk.Label(self.s3_table, text=h, style="ExpHeader.TLabel", padding=5, anchor='center', width=widths3[j]).grid(row=0, column=j, sticky='nsew')
            self.s3_table.grid_columnconfigure(j, weight=1 if j in [1, 4] else 0)
        enable_column_resize(self.s3_table, len(headers3))
            
        # Outsource defaults from master
        outsource_defaults = []
        if self.master_app:
            outsource_defaults = self.master_app.get_outsource_defaults()
        else:
            outsource_defaults = [("케이엔디이", "방사선투과검사", 0, 15000)]

        for cat, content, qty, price in outsource_defaults:
            self._add_row_s3(cat, content, qty, price)

        for _ in range(2): self._add_row_s3()

        # --- Section 4: Social Insurance ---
        s4_frame = ttk.Frame(self)
        s4_frame.pack(fill='x', pady=5)
        ttk.Label(s4_frame, text="(4) 4대 보험료", font=('Malgun Gothic', 10, 'bold')).pack(side='left', padx=5)
        
        insurance_table = ttk.Frame(self)
        insurance_table.pack(fill='x')
        headers4 = ["구분", "산출 기준", "산출 인건비", "단가(요율)", "사전원가가액"]
        for j, h in enumerate(headers4):
            ttk.Label(insurance_table, text=h, style="ExpHeader.TLabel", padding=5, anchor='center', width=widths[j] if j < len(widths) else 20).grid(row=0, column=j, sticky='nsew')
            insurance_table.grid_columnconfigure(j, weight=1 if j in [1, 4] else 0)
        enable_column_resize(insurance_table, len(headers4))
            
        ttk.Label(insurance_table, text="4대 보험료", relief='solid', padding=5, anchor='center').grid(row=1, column=0, sticky='nsew')
        ttk.Label(insurance_table, text="산출인건비 X 요율(2024.7.1 기준)", relief='solid', padding=5, anchor='w').grid(row=1, column=1, sticky='nsew')
        self.lbl_insurance_base = ttk.Label(insurance_table, text="₩ 0", relief='solid', padding=5, anchor='e')
        self.lbl_insurance_base.grid(row=1, column=2, sticky='nsew')
        ttk.Label(insurance_table, text="10.6661%", relief='solid', padding=5, anchor='center').grid(row=1, column=3, sticky='nsew')
        self.lbl_insurance_amount = ttk.Label(insurance_table, text="0", relief='solid', padding=5, anchor='e')
        self.lbl_insurance_amount.grid(row=1, column=4, sticky='nsew')

        # --- Section 5: Depreciation ---
        s5_frame = ttk.LabelFrame(self, text="(5) 감기상각비 (Depreciation)")
        s5_frame.pack(fill='x', pady=5)
        
        self.s5_table = ttk.Frame(s5_frame)
        self.s5_table.pack(fill='x')
        
        headers5 = ["장비명", "사양", "내용년수", "수량", "사용일수", "감기비/일", "사전원가가액"]
        for j, h in enumerate(headers5):
            ttk.Label(self.s5_table, text=h, style="ExpHeader.TLabel", padding=5, anchor='center', width=widths[j]).grid(row=0, column=j, sticky='nsew')
            self.s5_table.grid_columnconfigure(j, weight=1 if j in [1, 6] else 0)
        enable_column_resize(self.s5_table, len(headers5))
            
        defaults_s5 = [
            ("PAUT 장비", "", 5, 1, 0, 44444),
            ("PAUT SCANNER (MANUAL)", "", 5, 1, 0, 5556),
            ("PAUT SCANNER (COBRA)", "", 5, 1, 0, 16667),
            ("YOKE", "", 5, 1, 0, 222),
            ("현상용 탑차(5년간 보험비 포함)", "현장별 차량기입시 탑차 구분 기입", 5, 1, 0, 16667),
            ("스타렉스(5년간 보험비 포함)", "현장별 차량기입시 스타렉스 구분 기입", 5, 1, 0, 16667)
        ]
        for item, spec, life, qty, days, rate in defaults_s5:
            self._add_row_s5(item, spec, life, qty, days, rate)

        # --- TOTALS SUMMARY ---
        summary_frame = ttk.Frame(self, padding=10)
        summary_frame.pack(fill='x', pady=10)
        
        # Row: Expense Total (1~5)
        ttk.Label(summary_frame, text="경비 합계 : (1)~(5) 합계", style="ExpTotal.TLabel", anchor='center', padding=8).grid(row=0, column=0, columnspan=4, sticky='nsew')
        self.lbl_exp_total = ttk.Label(summary_frame, text="₩ 0", style="ExpTotal.TLabel", anchor='e', padding=8)
        self.lbl_exp_total.grid(row=0, column=4, sticky='nsew')
        
        # Row: Expense VAT
        ttk.Label(summary_frame, text="경비 부가세 합계", relief='solid', anchor='center', padding=5).grid(row=1, column=3, sticky='nsew')
        self.lbl_exp_vat = ttk.Label(summary_frame, text="0", relief='solid', anchor='e', padding=5)
        self.lbl_exp_vat.grid(row=1, column=4, sticky='nsew')

        # Row: Total Direct Cost (Sales Cost)
        ttk.Label(summary_frame, text="매출원가 총계 : 1), 2), 3) 합계", font=('Malgun Gothic', 10, 'bold'), background='#00ffff', relief='solid', anchor='center', padding=8).grid(row=2, column=0, columnspan=4, sticky='nsew')
        self.lbl_sales_cost_total = ttk.Label(summary_frame, text="₩ 0", font=('Malgun Gothic', 10, 'bold'), background='#00ffff', relief='solid', anchor='e', padding=8)
        self.lbl_sales_cost_total.grid(row=2, column=4, sticky='nsew')

        # Row: Indirect Cost (판관비)
        ttk.Label(summary_frame, text="3. 간접비(판관비)", font=('Malgun Gothic', 10, 'bold'), anchor='w').grid(row=3, column=0, pady=(10, 0))
        
        indirect_table = ttk.Frame(summary_frame)
        indirect_table.grid(row=4, column=0, columnspan=5, sticky='ew')
        headers_ind = ["구분", "산출 기준", "산출직접비", "간접비율(%)", "사전 간접비 합계"]
        for j, h in enumerate(headers_ind):
             ttk.Label(indirect_table, text=h, style="ExpHeader.TLabel", padding=5, anchor='center', width=widths[j] if j < len(widths) else 25).grid(row=0, column=j, sticky='nsew')
             indirect_table.grid_columnconfigure(j, weight=1 if j in [1, 4] else 0)
        enable_column_resize(indirect_table, len(headers_ind))
        
        ttk.Label(indirect_table, text="간접비", relief='solid', padding=5, anchor='center').grid(row=1, column=0, sticky='nsew')
        ttk.Label(indirect_table, text="산출직접비 x 간접비율(2024년 기준)", relief='solid', padding=5, anchor='w').grid(row=1, column=1, sticky='nsew')
        self.lbl_indirect_base = ttk.Label(indirect_table, text="₩ 0", relief='solid', padding=5, anchor='e')
        self.lbl_indirect_base.grid(row=1, column=2, sticky='nsew')
        ttk.Label(indirect_table, text="14%", relief='solid', padding=5, anchor='center').grid(row=1, column=3, sticky='nsew')
        self.lbl_indirect_total = ttk.Label(indirect_table, text="₩ 0", font=('Malgun Gothic', 10, 'bold'), background='#00ffff', relief='solid', anchor='e', padding=5)
        self.lbl_indirect_total.grid(row=1, column=4, sticky='nsew')

        # Row: Total Cost (Direct + Indirect)
        ttk.Label(summary_frame, text="4. 총원가(매출원가+간접비)", style="ExpTotal.TLabel", anchor='center', padding=10).grid(row=5, column=0, columnspan=4, sticky='nsew', pady=(10, 0))
        self.lbl_grand_total_cost = ttk.Label(summary_frame, text="₩ 0", style="ExpTotal.TLabel", anchor='e', padding=10)
        self.lbl_grand_total_cost.grid(row=5, column=4, sticky='nsew', pady=(10, 0))

        # Row: Operating Profit Section
        ttk.Label(summary_frame, text="5. 영업이익, 영업이익률", font=('Malgun Gothic', 10, 'bold'), anchor='w').grid(row=6, column=0, pady=(10, 0))
        
        profit_table = ttk.Frame(summary_frame)
        profit_table.grid(row=7, column=0, columnspan=5, sticky='ew')
        headers_prof = ["구분", "매출(수입)", "총원가", "영업이익", "영업이익률", "기준"]
        widths_prof = [15, 20, 20, 20, 15, 10]
        for j, h in enumerate(headers_prof):
             ttk.Label(profit_table, text=h, style="ExpHeader.TLabel", padding=5, anchor='center', width=widths_prof[j]).grid(row=0, column=j, sticky='nsew')
             profit_table.grid_columnconfigure(j, weight=1 if j != 5 else 0)
        enable_column_resize(profit_table, len(headers_prof))

        # Row 1: Budget (사전)
        ttk.Label(profit_table, text="사전(예산)", relief='solid', padding=5, anchor='center').grid(row=1, column=0, sticky='nsew')
        self.lbl_prof_revenue = ttk.Label(profit_table, text="₩ 0", relief='solid', padding=5, anchor='e')
        self.lbl_prof_revenue.grid(row=1, column=1, sticky='nsew')
        self.lbl_prof_total_cost = ttk.Label(profit_table, text="₩ 0", relief='solid', padding=5, anchor='e')
        self.lbl_prof_total_cost.grid(row=1, column=2, sticky='nsew')
        self.lbl_prof_op_profit = ttk.Label(profit_table, text="₩ 0", style="Margin.TLabel", padding=5, anchor='center')
        self.lbl_prof_op_profit.grid(row=1, column=3, sticky='nsew')
        self.lbl_prof_margin = ttk.Label(profit_table, text="0.00%", relief='solid', padding=5, anchor='center')
        self.lbl_prof_margin.grid(row=1, column=4, sticky='nsew')
        ttk.Label(profit_table, text="부가세 별도", relief='solid', padding=5, anchor='center').grid(row=1, column=5, sticky='nsew')

    def _add_row_s1(self, cat="", cont="", ppl="", qty="", unit="", price=0):
        row = len(self.entries['site_expense']) + 1
        widgets = {}
        ent_cat = ttk.Entry(self.s1_table, width=15, justify='center'); ent_cat.insert(0, cat); ent_cat.grid(row=row, column=0, sticky='nsew')
        ent_cont = ttk.Entry(self.s1_table, width=20); ent_cont.insert(0, cont); ent_cont.grid(row=row, column=1, sticky='nsew')
        ent_ppl = ttk.Entry(self.s1_table, width=8, justify='center'); ent_ppl.insert(0, str(ppl)); ent_ppl.grid(row=row, column=2, sticky='nsew')
        ent_qty = ttk.Entry(self.s1_table, width=8, justify='center'); ent_qty.insert(0, str(qty)); ent_qty.grid(row=row, column=3, sticky='nsew')
        ent_unit = ttk.Entry(self.s1_table, width=8, justify='center'); ent_unit.insert(0, unit); ent_unit.grid(row=row, column=4, sticky='nsew')
        ent_price = ttk.Entry(self.s1_table, width=15, justify='right'); ent_price.insert(0, f"{price:,.0f}"); ent_price.grid(row=row, column=5, sticky='nsew')
        lbl_amt = ttk.Label(self.s1_table, text="0", relief='solid', anchor='e', padding=5); lbl_amt.grid(row=row, column=6, sticky='nsew')
        
        widgets = {'cat': ent_cat, 'cont': ent_cont, 'ppl': ent_ppl, 'qty': ent_qty, 'unit': ent_unit, 'price': ent_price, 'amount': lbl_amt}
        for w in [ent_ppl, ent_qty, ent_price]: w.bind("<KeyRelease>", lambda e: self.calculate_all())
        self.entries['site_expense'].append(widgets)

    def _add_row_s2(self, cat="", spec="", qty="", period="", unit="", price=0):
        row = len(self.entries['rental']) + 1
        widgets = {}
        e1 = ttk.Entry(self.s2_table, width=15, justify='center'); e1.insert(0, cat); e1.grid(row=row, column=0, sticky='nsew')
        e2 = ttk.Entry(self.s2_table, width=20); e2.insert(0, spec); e2.grid(row=row, column=1, sticky='nsew')
        e3 = ttk.Entry(self.s2_table, width=8, justify='center'); e3.insert(0, str(qty)); e3.grid(row=row, column=2, sticky='nsew')
        e4 = ttk.Entry(self.s2_table, width=8, justify='center'); e4.insert(0, str(period)); e4.grid(row=row, column=3, sticky='nsew')
        e5 = ttk.Entry(self.s2_table, width=8, justify='center'); e5.insert(0, unit); e5.grid(row=row, column=4, sticky='nsew')
        e6 = ttk.Entry(self.s2_table, width=15, justify='right'); e6.insert(0, f"{price:,.0f}"); e6.grid(row=row, column=5, sticky='nsew')
        lbl = ttk.Label(self.s2_table, text="0", relief='solid', anchor='e', padding=5); lbl.grid(row=row, column=6, sticky='nsew')
        
        widgets = {'cat': e1, 'spec': e2, 'qty': e3, 'period': e4, 'unit': e5, 'price': e6, 'amount': lbl}
        for w in [e3, e4, e6]: w.bind("<KeyRelease>", lambda e: self.calculate_all())
        self.entries['rental'].append(widgets)

    def _add_row_s3(self, cat="", work="", count=0, price=0):
        row = len(self.entries['outsource']) + 1
        widgets = {}
        e1 = ttk.Entry(self.s3_table, width=15, justify='center'); e1.insert(0, cat); e1.grid(row=row, column=0, sticky='nsew')
        e2 = ttk.Entry(self.s3_table, width=30); e2.insert(0, work); e2.grid(row=row, column=1, sticky='nsew')
        e3 = ttk.Entry(self.s3_table, width=10, justify='center'); e3.insert(0, str(count)); e3.grid(row=row, column=2, sticky='nsew')
        e4 = ttk.Entry(self.s3_table, width=15, justify='right'); e4.insert(0, f"{price:,.0f}"); e4.grid(row=row, column=3, sticky='nsew')
        lbl = ttk.Label(self.s3_table, text="0", relief='solid', anchor='e', padding=5); lbl.grid(row=row, column=4, sticky='nsew')
        
        widgets = {'cat': e1, 'work': e2, 'count': e3, 'price': e4, 'amount': lbl}
        for w in [e3, e4]: w.bind("<KeyRelease>", lambda e: self.calculate_all())
        self.entries['outsource'].append(widgets)

    def _add_row_s5(self, item="", spec="", life=5, qty=1, days=0, rate=0):
        row = len(self.entries['depreciation']) + 1
        widgets = {}
        e1 = ttk.Entry(self.s5_table, width=20); e1.insert(0, item); e1.grid(row=row, column=0, sticky='nsew')
        e2 = ttk.Entry(self.s5_table, width=15); e2.insert(0, spec); e2.grid(row=row, column=1, sticky='nsew')
        e3 = ttk.Entry(self.s5_table, width=8, justify='center'); e3.insert(0, str(life)); e3.grid(row=row, column=2, sticky='nsew')
        e4 = ttk.Entry(self.s5_table, width=8, justify='center'); e4.insert(0, str(qty)); e4.grid(row=row, column=3, sticky='nsew')
        e5 = ttk.Entry(self.s5_table, width=8, justify='center'); e5.insert(0, str(days)); e5.grid(row=row, column=4, sticky='nsew')
        e6 = ttk.Entry(self.s5_table, width=15, justify='right'); e6.insert(0, f"{rate:,.0f}"); e6.grid(row=row, column=5, sticky='nsew')
        lbl = ttk.Label(self.s5_table, text="0", relief='solid', anchor='e', padding=5); lbl.grid(row=row, column=6, sticky='nsew')
        
        widgets = {'item': e1, 'spec': e2, 'life': e3, 'qty': e4, 'days': e5, 'rate': e6, 'amount': lbl}
        for w in [e4, e5, e6]: w.bind("<KeyRelease>", lambda e: self.calculate_all())
        self.entries['depreciation'].append(widgets)

    def calculate_all(self, event=None):
        # 1. Site Expenses
        t1 = 0.0
        for w in self.entries['site_expense']:
            amt = self._to_f(w['qty'].get()) * self._to_f(w['price'].get())
            w['amount'].config(text=f"{amt:,.0f}")
            t1 += amt
            
        # 2. Rentals
        t2 = 0.0
        for w in self.entries['rental']:
            amt = self._to_f(w['qty'].get()) * self._to_f(w['period'].get()) * self._to_f(w['price'].get())
            w['amount'].config(text=f"{amt:,.0f}")
            t2 += amt
            
        # 3. Outsource
        t3 = 0.0
        for w in self.entries['outsource']:
            amt = self._to_f(w['count'].get()) * self._to_f(w['price'].get())
            w['amount'].config(text=f"{amt:,.0f}")
            t3 += amt
            
        # 4. Insurance
        labor_total = self.get_labor_total() if self.get_labor_total else 0.0
        t4 = labor_total * 0.106661
        self.lbl_insurance_base.config(text=f"₩ {labor_total:,.0f}")
        self.lbl_insurance_amount.config(text=f"{t4:,.0f}")
        
        # 5. Depreciation
        t5 = 0.0
        for w in self.entries['depreciation']:
            amt = self._to_f(w['qty'].get()) * self._to_f(w['days'].get()) * self._to_f(w['rate'].get())
            w['amount'].config(text=f"{amt:,.0f}")
            t5 += amt
            
        exp_total = t1 + t2 + t3 + t4 + t5
        exp_vat = (t1 + t2 + t3) * 0.1 # Example: VAT on direct expenses
        self.lbl_exp_total.config(text=f"₩ {exp_total:,.0f}")
        self.lbl_exp_vat.config(text=f"{exp_vat:,.0f}")
        
        # 6. Sales Cost (Labor + Material + Exp)
        mat_total = self.get_material_total() if self.get_material_total else 0.0
        direct_cost = labor_total + mat_total + exp_total
        self.lbl_sales_cost_total.config(text=f"₩ {direct_cost:,.0f}")
        
        # 7. Indirect Cost (14%)
        self.lbl_indirect_base.config(text=f"₩ {direct_cost:,.0f}")
        indirect_cost = direct_cost * 0.14
        self.lbl_indirect_total.config(text=f"₩ {indirect_cost:,.0f}")
        
        grand_total_cost = direct_cost + indirect_cost
        self.lbl_grand_total_cost.config(text=f"₩ {grand_total_cost:,.0f}")
        
        # 8. Profit
        revenue = self.get_revenue() if self.get_revenue else 0.0
        op_profit = revenue - grand_total_cost
        margin = (op_profit / revenue * 100) if revenue > 0 else 0.0
        
        self.lbl_prof_revenue.config(text=f"₩ {revenue:,.0f}")
        self.lbl_prof_total_cost.config(text=f"₩ {grand_total_cost:,.0f}")
        self.lbl_prof_op_profit.config(text=f"₩ {op_profit:,.0f}")
        self.lbl_prof_margin.config(text=f"{margin:.2f}%")
        
        # Update main form "Expense" and "Outsource" fields
        if self.on_change_callback:
            # We pass (Expense, Outsource, TotalProfit) or something?
            # Let's just update the main Expense field with exp_total (1~5)
            # and Outsource with t3.
            self.on_change_callback(exp_total, t3, op_profit)

    def _to_f(self, val):
        try:
            return float(str(val).replace(',', '') or 0)
        except:
            return 0.0

    def get_total_cost(self):
        """[FINAL_FIX] Robustly get total expense cost (including depreciation)"""
        try:
            raw_text = self.lbl_exp_total.cget('text')
            val = "".join(c for c in raw_text if c.isdigit() or c == '.')
            return float(val or 0)
        except:
            return 0.0

    def get_data(self):
        data = {
            'site_expense': [{k: v.get() if hasattr(v, 'get') else v.cget('text') for k, v in row.items()} for row in self.entries['site_expense']],
            'rental': [{k: v.get() if hasattr(v, 'get') else v.cget('text') for k, v in row.items()} for row in self.entries['rental']],
            'outsource': [{k: v.get() if hasattr(v, 'get') else v.cget('text') for k, v in row.items()} for row in self.entries['outsource']],
            'depreciation': [{k: v.get() if hasattr(v, 'get') else v.cget('text') for k, v in row.items()} for row in self.entries['depreciation']]
        }
        return data

    def get_total_cost(self):
        """Retrieve the total site expense cost (items 1~5) as a float"""
        try:
            val = self.lbl_exp_total.cget('text').replace('₩', '').replace(',', '').replace(' ', '').strip()
            return float(val or 0)
        except:
            return 0.0

    def set_data(self, data):
        if not data or not isinstance(data, dict):
            self.reset()
            return
            
        def fill(entry_list, data_list):
            for i, d in enumerate(data_list):
                if i < len(entry_list):
                    for k, v in d.items():
                        if k in entry_list[i] and hasattr(entry_list[i][k], 'delete'):
                            entry_list[i][k].delete(0, tk.END); entry_list[i][k].insert(0, str(v))
        
        fill(self.entries['site_expense'], data.get('site_expense', []))
        fill(self.entries['rental'], data.get('rental', []))
        fill(self.entries['outsource'], data.get('outsource', []))
        fill(self.entries['depreciation'], data.get('depreciation', []))
        
        self.calculate_all()

    def reset(self):
        def clear(entry_list):
            for row in entry_list:
                for k, v in row.items():
                    if hasattr(v, 'delete'): v.delete(0, tk.END)
        
        clear(self.entries['site_expense'])
        clear(self.entries['rental'])
        clear(self.entries['outsource'])
        clear(self.entries['depreciation'])
        self.calculate_all()

class ColumnSelectionDialog(tk.Toplevel):
    """Dialog to select columns for Excel export"""
    def __init__(self, parent, columns, title="엑셀 출력 컬럼 선택"):
        super().__init__(parent)
        self.title(title)
        self.geometry("500x700") # Larger for better visibility
        self.transient(parent)
        self.grab_set()
        
        self.result = None
        self.vars = {}
        
        # Header
        lbl = ttk.Label(self, text="출력할 컬럼을 선택하세요:", font=('Malgun Gothic', 11, 'bold'))
        lbl.pack(pady=10)
        
        # Scrollable area for checkboxes
        container = ttk.Frame(self)
        container.pack(fill='both', expand=True, padx=20)
        
        canvas = tk.Canvas(container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Unbind when closed to avoid errors
        def on_close():
            self.destroy()
        
        self.protocol("WM_DELETE_WINDOW", on_close)
            
        # Create checkboxes
        for item in columns:
            if isinstance(item, (tuple, list)) and len(item) >= 2:
                col = item[0]
                provided_display_text = item[1]
            else:
                col = item
                provided_display_text = None
            var = tk.BooleanVar(value=True)
            self.vars[col] = var
            # 사용자에게는 내부 컬럼명 대신 표시용 한글명을 보여줌
            display_map = {
                '검사량': '수량',
                'Date': '날짜',
                'Site': '현장',
                'EntryTime': '입력시간',
                'MaterialID': '자재ID',
                'FilmCount': '수량'
            }
            display_text = provided_display_text or display_map.get(col, col)
            cb = ttk.Checkbutton(scrollable_frame, text=display_text, variable=var)
            cb.pack(anchor='w', pady=2)
            
        # Buttons area
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill='x', pady=15, padx=20)
        
        def select_all():
            for v in self.vars.values(): v.set(True)
        def deselect_all():
            for v in self.vars.values(): v.set(False)
            
        ttk.Button(btn_frame, text="전체 선택", command=select_all).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="전체 해제", command=deselect_all).pack(side='left', padx=5)
        
        # Bottom controls
        bottom_frame = ttk.Frame(self)
        bottom_frame.pack(fill='x', pady=10)
        
        def on_ok():
            canvas.unbind_all("<MouseWheel>") # Unbind on OK too
            self.result = [col for col, var in self.vars.items() if var.get()]
            self.destroy()
            
        def on_cancel():
            on_close()
            
        ttk.Button(bottom_frame, text="확인", command=on_ok).pack(side='right', padx=10)
        ttk.Button(bottom_frame, text="취소", command=on_close).pack(side='right', padx=5)
        
        # self.wait_window() removed from here to allow caller to set vars before blocking
        # self.wait_window() removed from here to allow caller to set vars before blocking

class MaterialManager:
    def __init__(self, root):
        self.root = root
        
        # High DPI awareness
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            try:
                ctypes.windll.user32.SetProcessDPIAware()
            except Exception:
                pass
                
        self.root.title("자재 및 소모품 관리 시스템 (Material Manager)")
        self.root.geometry("1600x900")
        try:
            self.root.state('zoomed') # Maximize on Windows
        except:
            pass
        
        # Configure overall style
        
        # Global Scroll Handler for all canvases in the app
        def _on_global_mousewheel(event):
            try:
                widget = event.widget.winfo_containing(event.x_root, event.y_root)
                if not widget: return
                
                # Do not interfere with native scrolling widgets
                if isinstance(widget, (tk.Text, ttk.Treeview, tk.Listbox)):
                    return
                    
                parent = widget
                while parent:
                    if isinstance(parent, tk.Canvas):
                        # Allow scrolling on all canvases
                        parent.yview_scroll(int(-1 * (event.delta / 120)), "units")
                        return "break" # [FIX] Prevent scroll from reaching Combobox/Entry under mouse

                    if hasattr(parent, 'master') and parent.master:
                        parent = parent.master
                    else:
                        break
            except Exception:
                pass
            return "break" # Default to break if we are over a canvas-descendant to be safe
                
        self.root.bind_all("<MouseWheel>", _on_global_mousewheel)
        
        # [FIX] Globally disable MouseWheel on Combobox and Entry to prevent unintentional value changes during scrolling
        self.root.bind_class("TCombobox", "<MouseWheel>", lambda e: "break")
        self.root.bind_class("TEntry", "<MouseWheel>", lambda e: "break")
        
        self.style = ttk.Style()
        try:
            self.style.theme_use('clam') # Use 'clam' theme for better grid line visibility
        except:
            pass
            
        self.style.configure(".", font=('Malgun Gothic', 12))
        self.style.configure("Treeview.Heading", font=('Malgun Gothic', 12, 'bold'))
        self.style.configure("Treeview", font=('Malgun Gothic', 12), rowheight=35) # Increased row height for "boxed" look
        
        # Detect system background color for tk widgets (Canvas, Text)
        self.theme_bg = self.root.cget('bg')
        if not self.theme_bg or self.theme_bg == 'SystemButtonFace':
             # Fallback/Modern check for Windows
             try:
                 self.theme_bg = self.style.lookup('TFrame', 'background')
             except:
                 self.theme_bg = '#f0f0f0' # Standard Windows gray
        
        # Detachable Windows State (pop-up support)
        self.detached_windows = {}
        
        # Determine base directory and bundle directory for portability
        if getattr(sys, 'frozen', False):
            # If running as an executable
            self.app_dir = os.path.dirname(sys.executable)
            self.bundle_dir = getattr(sys, '_MEIPASS', self.app_dir)
            # For exe, use user's Documents folder for config to ensure write permissions
            documents_dir = os.path.join(os.path.expanduser('~'), 'Documents', 'MaterialManager')
            if not os.path.exists(documents_dir):
                os.makedirs(documents_dir, exist_ok=True)
            self.config_path = os.path.join(documents_dir, 'Material_Manager_Config.json')
        else:
            # 스크립트 실행 모드: src/ 우선, 없으면 ../data/ 탐색
            self.app_dir = os.path.dirname(os.path.abspath(__file__))
            self.bundle_dir = self.app_dir
            self.config_path = os.path.join(self.app_dir, 'Material_Manager_Config.json')
            _db_name = 'Material_Inventory.xlsx'
            _data_dir = os.path.join(os.path.dirname(self.app_dir), 'data')
            _candidates = [
                os.path.join(self.app_dir, _db_name),   # src/Material_Inventory.xlsx
                os.path.join(_data_dir, _db_name),      # data/Material_Inventory.xlsx
            ]
            self.db_path = next((p for p in _candidates if os.path.exists(p)), _candidates[0])

        if getattr(sys, 'frozen', False):
            # exe 실행 모드: exe와 같은 폴더에 DB 저장 (어디서 실행해도 동일)
            self.db_path = os.path.join(self.app_dir, 'Material_Inventory.xlsx')

        # db_path 확정
        
        self.sites = [] # Initialize site list
        self.daily_units = ['EA', 'CAN', 'SET', 'KG', 'M', '매', 'I/D', 'P,M,I/D', 'M,I/D', 'Point', 'Meter', 'Inch', 'Dia']
        self.users = [
            "부장 주진철", "대리 우명광", "주임 김진환", "계장 장승대", "주임 김성렬", "부장 박광복", "과장 주영광"
        ] # Initialize worker/name list
        self.warehouses = [] # Initialize warehouse list
        self.equipments = [] # Initialize equipment list
        self.vehicles = [] # Initialize vehicle list
        self.vehicle_inspections = {} # [NEW] Track active inspection widgets
        self.vehicle_boxes = [] # [NEW] Track active inspection widgets for automated save
        self.companies = [] # Initialize companies list
        self.worktimes = [] # Initialize worktimes list
        self.ot_times = [] # Initialize ot_times list
        
        # [NEW] Pre-initialize autocomplete/display lists to prevent AttributeError during load_data/refresh_filters
        self.materials_display_list = []
        self.co_code_list = []
        self.eq_code_list = []
        self.item_name_list = []
        self.class_list = []
        self.spec_list = []
        self.unit_list = []
        self.mfr_list = []
        self.origin_list = []
        self.sn_list = []
        self.model_list = []
        self.supplier_list = []
        self.equipment_suggestions = []
        self.test_methods = ["RT", "PAUT", "UT", "MT", "PT", "PMI"]  # [NEW] Initialize test methods
        
        # [NEW] Centralized NDT Consumable Definitions
        self.ndt_groups = {
            'PT약품': ['세척제', '침투제', '현상제', '형광침투제'],
            'MT약품': ['백색페인트', '흑색자분', '형광자분']
        }
        self.ndt_materials_all = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
        
        self.budget_sites = []  # [NEW] 공사실행예산서 전용 현장 목록 (budget_df 현장만)
        self.hidden_sites = ["초안", "롯데현장"]  # [NEW] Default hidden sites as requested
        
        self.load_data()
        
        # Centralized list of Carestream films for suggestions
        self.carestream_films = [
            "Carestream AA400-3⅓*12\"",
            "Carestream AA400-3⅓*17\"", 
            "Carestream AA400-4½*12\"",
            "Carestream AA400-10*12\"",
            "Carestream M100-3⅓*12\"",
            "Carestream M100-10*12\"",
            "Carestream M100-14*17\"",
            "Carestream MX125-3⅓*6\"",
            "Carestream MX125-3⅓*12\"",
            "Carestream MX125-4½*12\"",
            "Carestream MX125-10*12\"",
            "Carestream MX125-14*17\"",
            "Carestream T200-3⅓*12\"",
            "Carestream T200-3⅓*17\"",
            "Carestream T200-4½*12\"",
            "Carestream T200-10*12\"",
            "Carestream T200-14*17\""
        ]
        
        # Core draggable widget keys that should never be hidden
        self.CORE_DRAGGABLE_KEYS = [
            'form_box_geometry', 
            'ndt_usage_box_geometry', 
            'rtk_usage_box_geometry', 
            'save_btn_geometry', 
            'workers_box_geometry'
        ]
        
        # Registry for draggable items {config_key: widget_instance}
        self.draggable_items = {}
        self.memos = {} # key -> {'container': container, 'text_widget': text, 'title_entry': entry}
        self.checklists = {} # key -> {'container': container, 'title_entry': entry, 'item_frame': frame, 'items': []}
        self.vehicle_inspections = {} # key -> widget
        self.layout_locked = False
        self.daily_usage_sash_locked = False
        self._last_motion_time = 0 # For performance throttling
        self.is_ready = False  # Suppress saves until fully loaded

        # [CRITICAL FIX] Pre-load configuration synchronously to ensure UI starts in correct state
        self.preload_config_locks()

        self.create_widgets()
        self.update_registration_combos()
        
        # Enable keyboard navigation
        self.setup_keyboard_shortcuts()

    def preload_config_locks(self):
        """Pre-load critical lock states synchronously before UI creation"""
        try:
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.layout_locked = config.get('layout_locked', False)
                    self.daily_usage_sash_locked = config.get('daily_usage_sash_locked', False)
                    print(f"DEBUG: Pre-loaded locks - Layout: {self.layout_locked}, Sash: {self.daily_usage_sash_locked}")
        except Exception as e:
            print(f"DEBUG: Failed to pre-load locks: {e}")

    def _ensure_canvas_scroll_region(self):
        """Update canvas scroll region based on content height and width (ensures full visibility)"""
        try:
            if hasattr(self, 'entry_canvas') and self.entry_canvas:
                self.entry_canvas.update_idletasks()
                
                # Get max Y and X from all core elements
                max_y = 0
                max_x = 1100 # Default minimum width
                
                # 1. Use the bounding box of master_form_panel which contains form, NDT, RTK, and Workers
                if hasattr(self, 'master_form_panel'):
                    self.master_form_panel.update_idletasks()
                    panel_y = self.master_form_panel.winfo_y()
                    panel_h = self.master_form_panel.winfo_height()
                    panel_x = self.master_form_panel.winfo_x()
                    panel_w = self.master_form_panel.winfo_width()
                    max_y = max(max_y, panel_y + panel_h)
                    max_x = max(max_x, panel_x + panel_w)
                
                # 2. Specifically check RTK bottom if requested by user
                if hasattr(self, 'rtk_grid'):
                    # rtk_grid is inside master_form_panel, so calculate relative to master_form_panel master
                    self.rtk_grid.update_idletasks()
                    rtk_bottom = self.rtk_grid.winfo_y() + self.rtk_grid.winfo_height()
                    # Add master_form_panel offset
                    if hasattr(self, 'master_form_panel'):
                        rtk_bottom += self.master_form_panel.winfo_y()
                    max_y = max(max_y, rtk_bottom)
                
                # 3. Handle draggable items (Memos, Checklists, etc.)
                for key, widget in self.draggable_items.items():
                    try:
                        if widget.winfo_manager() == 'place':
                            info = widget.place_info()
                            x = int(float(info.get('x', 0)))
                            y = int(float(info.get('y', 0)))
                            w = int(float(info.get('width', widget.winfo_width())))
                            h = int(float(info.get('height', widget.winfo_height())))
                            max_y = max(max_y, y + h)
                            max_x = max(max_x, x + w)
                    except: pass

                # Final scroll height and width with minimal buffer
                scroll_h = max_y + 10
                scroll_w = max(max_x, self.entry_inner_frame.winfo_width())
                
                self.entry_canvas.configure(scrollregion=(0, 0, scroll_w, scroll_h))
        except Exception as e:
            print(f"DEBUG: Scroll region update error: {e}")

    def _on_daily_usage_sash_changed(self, event=None):
        """Handle sash position change to save ratio"""
        try:
            # Skip saving if locked
            if hasattr(self, 'daily_usage_sash_locked') and self.daily_usage_sash_locked:
                return

            if hasattr(self, 'daily_usage_paned'):
                self.daily_usage_paned.update_idletasks()
                total_h = self.daily_usage_paned.winfo_height()
                if total_h > 0:
                    sash_pos = self.daily_usage_paned.sashpos(0)
                    ratio = sash_pos / total_h
                    
                    if not hasattr(self, 'tab_config'):
                        self.tab_config = {}
                    
                    self.tab_config['daily_usage_sash_ratio'] = ratio
                    self.tab_config['daily_usage_sash_pos'] = sash_pos
                    
                    self.save_tab_config()
                    print(f"Sash ratio saved: {ratio:.3f}")
        except Exception as e:
            print(f"Error saving sash position: {e}")

    def toggle_resolution_lock(self):
        """Toggle window resolution lock"""
        try:
            self.resolution_locked = not self.resolution_locked
            
            if self.resolution_locked:
                self.locked_width = self.root.winfo_width()
                self.locked_height = self.root.winfo_height()
                self.root.resizable(False, False)
                if hasattr(self, 'btn_resolution_lock'):
                    self.btn_resolution_lock.config(text="🔒 해상도 고정됨")
                
                if not hasattr(self, 'tab_config'):
                    self.tab_config = {}
                self.tab_config['resolution_locked'] = True
                self.tab_config['locked_width'] = self.locked_width
                self.tab_config['locked_height'] = self.locked_height
                print(f"Resolution locked at: {self.locked_width}x{self.locked_height}")
            else:
                self.root.resizable(True, True)
                if hasattr(self, 'btn_resolution_lock'):
                    self.btn_resolution_lock.config(text="🔓 해상도 고정")
                
                if not hasattr(self, 'tab_config'):
                    self.tab_config = {}
                self.tab_config['resolution_locked'] = False
                print("Resolution unlocked")
                
            self.save_tab_config(force=True)
        except Exception as e:
            print(f"Error toggling resolution lock: {e}")

    # Removed force_save_config - using save_tab_config(force=True) instead

    def extract_sn_from_model(self, model_name, current_sn):
        """Extract SN from model name if 'S/N.' is present and update current_sn if it's empty"""
        if pd.isna(model_name):
            return model_name, current_sn
        
        model_str = str(model_name)
        if 'S/N.' in model_str:
            parts = model_str.split('S/N.', 1)
            new_model = parts[0].strip()
            # Remove trailing underscore or dot if present before S/N.
            if new_model.endswith('_') or new_model.endswith('.'):
                new_model = new_model[:-1].strip()
            
            extracted_sn = parts[1].strip()
            
            # Use extracted SN if current SN is empty or matches part of it
            if not str(current_sn).strip() or str(current_sn) == 'nan':
                return new_model, extracted_sn
            else:
                # If current SN exists, we still clean up the model name
                return new_model, current_sn
        
        return model_name, current_sn

    def _sync_dataframe_schema(self, df, sheet_name):
        """Ensure the given DataFrame has all required columns for its sheet_name."""
        if df is None:
            return None
            
        schemas = {
            'Materials': [
                'MaterialID', '회사코드', '관리품번', '품목명', 'SN', '창고',
                '모델명', '규격', '품목군코드', '공급업체', '제조사', '제조국', 
                '가격', '원가', '관리단위', '수량', '재고하한', 'Active'
            ],
            'Transactions': [
                'Date', 'MaterialID', 'Site', 'Type', 'Quantity', 'Note', 'User', 
                '차량번호', '주행거리', '차량점검', '차량비고'
            ],
            'MonthlyUsage': [
                'MaterialID', 'Year', 'Month', 'Site', 'Usage', 'Note', 'EntryDate'
            ],
            'DailyUsage': [
                'Date', 'Site', '업체명', 'MaterialID', 'Usage', 'Note', 'EntryTime',
                'RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크',
                'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타', '장비명', '검사방법', '검사량',
                '단가', '출장비', '일식', '검사비', '회사코드', 'FilmCount',
                'User', 'WorkTime', 'OT',
                'User2', 'WorkTime2', 'OT2',
                'User3', 'WorkTime3', 'OT3',
                'User4', 'WorkTime4', 'OT4',
                'User5', 'WorkTime5', 'OT5',
                'User6', 'WorkTime6', 'OT6',
                'User7', 'WorkTime7', 'OT7',
                'User8', 'WorkTime8', 'OT8',
                'User9', 'WorkTime9', 'OT9',
                'User10', 'WorkTime10', 'OT10',
                '차량번호', '주행거리', '차량점검', '차량비고'
            ],
            'Budget': [
                'Site', 'Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 
                'Expense', 'OutsourceCost', 'Profit', 'Note', 'LaborDetail', 'MaterialDetail'
            ]
        }
        
        required_cols = schemas.get(sheet_name, [])
        any_added = False
        for col in required_cols:
            if col not in df.columns:
                # Default values based on col type heuristics
                if col in ['Quantity', 'Usage', '검사량', '단가', '출장비', '일식', '검사비', 'FilmCount', '수량', '원가', '가격', '재고하한']:
                    df[col] = 0.0
                elif col == 'Active':
                    df[col] = 1
                elif col in ['Date', 'EntryTime', 'EntryDate', 'Year', 'Month']:
                    if col in ['Year', 'Month']: df[col] = 0
                    else: df[col] = pd.NaT
                else:
                    df[col] = ""
                any_added = True
        
        if any_added:
            print(f"DEBUG: Synchronized schema for {sheet_name}. Missing columns added.")
        return df

    def load_data(self):
        import re
        def normalize_cols(df):
            if df is not None and not df.empty:
                # 1. Standardize whitespace and handle numeric column names
                df.columns = [re.sub(r'\s+', '', str(c)) for c in df.columns]
                
                # 2. [ROBUST] Fallback mapping for common garbled column names
                # Instead of a complex map, we focus on the most critical ones using partial matches
                new_cols = {}
                for col in df.columns:
                    c = str(col)
                    if 'ǰ' in c: new_cols[col] = '품목명'
                    elif '𵨸' in c: new_cols[col] = '모델명'
                    elif 'â' in c: new_cols[col] = '창고'
                    elif '˻' in c: 
                        if '緮' in c or '量' in c: new_cols[col] = '검사량'
                        else: new_cols[col] = '검사방법'
                    elif 'ܰ' in c: new_cols[col] = '단가'
                    elif 'ȸڵ' in c: new_cols[col] = '회사코드'
                
                if new_cols:
                    df.rename(columns=new_cols, inplace=True)
            return df

        try:
            print(f"DEBUG: Loading data from {self.db_path}...")
            
            # Check if database exists in app_dir. If not, try to restore from bundle_dir
            if not os.path.exists(self.db_path):
                bundled_db = os.path.join(self.bundle_dir, 'Material_Inventory.xlsx')
                print(f"DEBUG: Main DB not found. Trying to restore from bundle: {bundled_db}")
                if os.path.exists(bundled_db):
                    import shutil
                    try:
                        shutil.copy2(bundled_db, self.db_path)
                        print("DEBUG: Restored DB from bundle.")
                        # Also try to copy config if it exists in bundle but not in app_dir
                        bundled_config = os.path.join(self.bundle_dir, 'Material_Manager_Config.json')
                        if os.path.exists(bundled_config) and not os.path.exists(self.config_path):
                            shutil.copy2(bundled_config, self.config_path)
                            print("DEBUG: Restored Config from bundle.")
                    except Exception as e:
                        print(f"Failed to restore data from bundle: {e}")

            if not os.path.exists(self.db_path):
                print("DEBUG: DB still not found. Initializing new DataFrames.")
                # Initialize with new schema if still not found
                self.materials_df = pd.DataFrame(columns=[
                    'MaterialID', '회사코드', '관리품번', '품목명', 'SN', '창고',
                    '모델명', '규격', '품목군코드', '공급업체', '제조사', '제조국', 
                    '가격', '원가', '관리단위', '수량', '재고하한', 'Active'
                ])
                self.transactions_df = pd.DataFrame(columns=['Date', 'MaterialID', 'Site', 'Type', 'Quantity', 'Note', 'User', '차량번호', '주행거리', '차량점검', '차량비고'])
                self.monthly_usage_df = pd.DataFrame(columns=['MaterialID', 'Year', 'Month', 'Site', 'Usage', 'Note', 'Entry Date'])
                self.daily_usage_df = pd.DataFrame(columns=['Date', 'Site', '업체명', 'MaterialID', 'Usage', 'Note', 'EntryTime',
                                                'RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크',
                                                'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타', '장비명', '검사방법', '검사량',
                                                '단가', '출장비', '일식', '검사비', '회사코드',
                                                'User', 'WorkTime', 'OT',
                                                'User2', 'WorkTime2', 'OT2',
                                                'User3', 'WorkTime3', 'OT3',
                                                'User4', 'WorkTime4', 'OT4',
                                                'User5', 'WorkTime5', 'OT5',
                                                'User6', 'WorkTime6', 'OT6',
                                                'User7', 'WorkTime7', 'OT7',
                                                'User8', 'WorkTime8', 'OT8',
                                                'User9', 'WorkTime9', 'OT9',
                                                'User10', 'WorkTime10', 'OT10',
                                                '차량번호', '주행거리', '차량점검', '차량비고'])
                self.budget_df = pd.DataFrame(columns=['Site', 'Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit', 'Note', 'LaborDetail', 'MaterialDetail'])
            else:
                print("DEBUG: DB found. Reading Excel...")
                # 1. Materials
                self.materials_df = pd.read_excel(self.db_path, sheet_name='Materials')
                self.materials_df = normalize_cols(self.materials_df)
                self.materials_df = self._sync_dataframe_schema(self.materials_df, 'Materials')
                
                # Handle column rename from '품명' to '품목명'
                if '품명' in self.materials_df.columns and '품목명' not in self.materials_df.columns:
                    self.materials_df.rename(columns={'품명': '품목명'}, inplace=True)
                
                # [NEW] Handle various quantity column aliases to prevent 0.0 readings
                qty_aliases = {'기초재고': '수량', '현재고': '수량', 'Stock': '수량', 'Quantity': '수량', 'Qty': '수량', '재고': '수량'}
                for alias, target in qty_aliases.items():
                    if alias in self.materials_df.columns and target not in self.materials_df.columns:
                        print(f"DEBUG: Mapping {alias} to {target}")
                        self.materials_df.rename(columns={alias: target}, inplace=True)
                
                print(f"DEBUG: Materials columns loaded: {self.materials_df.columns.tolist()}")
                
                # Ensure Active is numeric and handle NaNs (treat as Active=1)
                self.materials_df['Active'] = pd.to_numeric(self.materials_df['Active'], errors='coerce').fillna(1)
                
                # [STABILITY] Ensure MaterialID is consistently numeric
                if 'MaterialID' in self.materials_df.columns:
                    self.materials_df['MaterialID'] = pd.to_numeric(self.materials_df['MaterialID'], errors='coerce')
                
                # Force specific columns to string type and clean numeric artifacts (.0, -0.0)
                str_cols = ['회사코드', '관리품번', '품목명', 'SN', '창고', '모델명', '규격', 
                            '품목군코드', '공급업체', '제조사', '제조국', '관리단위']
                for col in str_cols:
                    if col in self.materials_df.columns:
                        # Ensure all strings are stripped and nan-free
                        self.materials_df[col] = self.materials_df[col].astype(str).str.strip().replace(['nan', 'None', 'NULL', '-0.0', '0.0', 'NaN', 'NaN.0'], '')
                        self.materials_df[col] = self.materials_df[col].str.replace(r'\.0$', '', regex=True)
                
                # [NEW] One-time Data Migration: Strip "MT " and "PT " prefixes from NDT medicine models
                # This ensures existing records are correctly summarized in the Inventory Status tab.
                if not self.materials_df.empty:
                    ndt_parent_cats = ["PT약품", "MT약품", "NDT약품"]
                    # Normalize category names for matching
                    temp_cats = self.materials_df['품목명'].str.replace(' ', '').str.upper()
                    mask = temp_cats.isin(ndt_parent_cats)
                    if mask.any():
                        def strip_ndt_prefix(model):
                            s = str(model).strip()
                            # Check for PT or MT prefix followed by a space
                            if s.upper().startswith("MT "): return s[3:].strip()
                            if s.upper().startswith("PT "): return s[3:].strip()
                            return s
                        
                        self.materials_df.loc[mask, '모델명'] = self.materials_df.loc[mask, '모델명'].apply(strip_ndt_prefix)
                        print(f"DEBUG: Migrated {mask.sum()} NDT items by stripping prefixes.")
                
                # 2. Transactions
                self.transactions_df = pd.read_excel(self.db_path, sheet_name='Transactions')
                self.transactions_df = normalize_cols(self.transactions_df)
                self.transactions_df = self._sync_dataframe_schema(self.transactions_df, 'Transactions')
                
                # Ensure Date column is datetime and MaterialID is numeric
                if not self.transactions_df.empty:
                    self.transactions_df['Date'] = pd.to_datetime(self.transactions_df['Date'], errors='coerce')
                    self.transactions_df['MaterialID'] = pd.to_numeric(self.transactions_df['MaterialID'], errors='coerce')
                    
                    # Force string columns and clean numeric artifacts
                    for col in ['Type', 'Note', 'User', 'Site', '차량번호', '주행거리', '차량점검', '차량비고']:
                        self.transactions_df[col] = self.transactions_df[col].astype(str).replace(['nan', 'None', 'NULL', '-0.0', '0.0', 'NaN'], '')
                        self.transactions_df[col] = self.transactions_df[col].str.replace(r'\.0$', '', regex=True)
                    
                    # One-time cleanup: Remove '현장사용' and redundant model names from historical notes
                    self.transactions_df['Note'] = self.transactions_df['Note'].astype(str).str.replace('현장사용', '', regex=False).str.strip()
                    
                    # Clean up notes that are identical to model names
                    if not self.transactions_df.empty and not self.materials_df.empty:
                        # Create a map for MaterialID -> Model Name
                        id_to_model = self.materials_df.set_index('MaterialID')['모델명'].astype(str).to_dict()
                        
                        def clean_redundant_note(row):
                            note = str(row['Note']).strip()
                            mat_id = row['MaterialID']
                            model = str(id_to_model.get(mat_id, '')).strip()
                            if note and model and note == model:
                                return ''
                            return note
                            
                        self.transactions_df['Note'] = self.transactions_df.apply(clean_redundant_note, axis=1)
                    
                    # Normalize OUT transaction quantities to negative and strip Site names
                    if not self.transactions_df.empty:
                        if 'Site' in self.transactions_df.columns:
                            self.transactions_df['Site'] = self.transactions_df['Site'].astype(str).str.strip().replace(['nan', 'None'], '')
                        
                        out_mask = (self.transactions_df['Type'] == 'OUT') & (self.transactions_df['Quantity'] > 0)
                        if out_mask.any():
                            self.transactions_df.loc[out_mask, 'Quantity'] = -self.transactions_df.loc[out_mask, 'Quantity']
                            print(f"DEBUG: Normalized {out_mask.sum()} OUT transactions to negative.")
                
                # 3. Monthly Usage
                try:
                    self.monthly_usage_df = pd.read_excel(self.db_path, sheet_name='MonthlyUsage', dtype={'Site': str, 'Note': str})
                    self.monthly_usage_df = normalize_cols(self.monthly_usage_df)
                    self.monthly_usage_df = self._sync_dataframe_schema(self.monthly_usage_df, 'MonthlyUsage')
                    
                    if not self.monthly_usage_df.empty:
                        self.monthly_usage_df['MaterialID'] = pd.to_numeric(self.monthly_usage_df['MaterialID'], errors='coerce')
                        self.monthly_usage_df['EntryDate'] = pd.to_datetime(self.monthly_usage_df['EntryDate'])
                        
                        self.monthly_usage_df['Site'] = self.monthly_usage_df['Site'].astype(str).str.strip().replace(['nan', 'None'], '')
                except Exception as e:
                    print(f"DEBUG: Failed to load MonthlyUsage: {e}")
                    self.monthly_usage_df = pd.DataFrame(columns=['MaterialID', 'Year', 'Month', 'Site', 'Usage', 'Note', 'EntryDate'])
                
                # 4. Daily Usage
                try:
                    # Explicitly set dtypes for vehicle and note columns to avoid float inference for empty cells
                    self.daily_usage_df = pd.read_excel(self.db_path, sheet_name='DailyUsage', 
                                                        dtype={'Site': str, 'Note': str, 'User': str,
                                                               '차량번호': str, '주행거리': str, '차량점검': str, '차량비고': str})
                    print(f"DEBUG: Loaded {len(self.daily_usage_df)} records from DailyUsage sheet.")
                    self.daily_usage_df = normalize_cols(self.daily_usage_df)
                    
                    # [NEW] Column Name Migration for Daily Usage
                    if '날짜' in self.daily_usage_df.columns:
                        if 'Date' not in self.daily_usage_df.columns:
                            self.daily_usage_df.rename(columns={'날짜': 'Date'}, inplace=True)
                        else:
                            # Merge if both exist
                            self.daily_usage_df['Date'] = self.daily_usage_df['Date'].fillna(self.daily_usage_df['날짜'])
                            self.daily_usage_df.drop(columns=['날짜'], inplace=True)

                    if '현장' in self.daily_usage_df.columns:
                        if 'Site' not in self.daily_usage_df.columns:
                            self.daily_usage_df.rename(columns={'현장': 'Site'}, inplace=True)
                        else:
                            self.daily_usage_df['Site'] = self.daily_usage_df['Site'].fillna(self.daily_usage_df['현장'])
                            self.daily_usage_df.drop(columns=['현장'], inplace=True)

                    if '수량' in self.daily_usage_df.columns:
                        if 'Usage' not in self.daily_usage_df.columns:
                            self.daily_usage_df.rename(columns={'수량': 'Usage'}, inplace=True)
                        else:
                            self.daily_usage_df['Usage'] = self.daily_usage_df['Usage'].fillna(self.daily_usage_df['수량'])
                            self.daily_usage_df.drop(columns=['수량'], inplace=True)

                    self.daily_usage_df = self._sync_dataframe_schema(self.daily_usage_df, 'DailyUsage')

                    # [NEW] Migrate FilmCount to Usage/수량 if needed (for legacy data)
                    if 'FilmCount' in self.daily_usage_df.columns or '필름매수' in self.daily_usage_df.columns:
                        f_col = 'FilmCount' if 'FilmCount' in self.daily_usage_df.columns else '필름매수'
                        u_col = 'Usage' if 'Usage' in self.daily_usage_df.columns else ('수량' if '수량' in self.daily_usage_df.columns else None)
                        
                        if u_col and f_col:
                            # If usage is 0/empty, take FilmCount as fallback for legacy records
                            self.daily_usage_df[u_col] = pd.to_numeric(self.daily_usage_df[u_col], errors='coerce').fillna(0)
                            self.daily_usage_df[f_col] = pd.to_numeric(self.daily_usage_df[f_col], errors='coerce').fillna(0)
                            mask = (self.daily_usage_df[u_col] == 0) & (self.daily_usage_df[f_col] > 0)
                            # self.daily_usage_df.loc[mask, u_col] = self.daily_usage_df.loc[mask, f_col]
                            # [DEFENSIVE] Do NOT migrate if we want to keep them separate now. 
                            # We just keep both columns.
                            pass 

                                                        
                    if not self.daily_usage_df.empty:
                        self.daily_usage_df['Date'] = pd.to_datetime(self.daily_usage_df['Date'])
                        self.daily_usage_df['EntryTime'] = pd.to_datetime(self.daily_usage_df['EntryTime'])
                        
                        # Fill NaNs and clean numeric artifacts in string columns
                        string_columns = ['Site', 'Note', '장비명', '검사방법', '업체명', '차량번호', '주행거리', '차량점검', '차량비고']
                        # Add all users, worktimes and OT columns
                        for i in range(1, 11):
                            u_col = 'User' if i == 1 else f'User{i}'
                            w_col = 'WorkTime' if i == 1 else f'WorkTime{i}'
                            o_col = 'OT' if i == 1 else f'OT{i}'
                            string_columns.extend([u_col, w_col, o_col])
                            
                        for col in string_columns:
                            if col not in self.daily_usage_df.columns:
                                self.daily_usage_df[col] = ''
                            
                            # [SAFE] Ensure NO numeric-locked columns receive an empty string by accident
                            # Only apply string replacement if the column is NOT designated as numeric later
                            numeric_intended = ['Usage', '검사량', '단가', '출장비', '일식', '검사비', '수량', 'OT', 'OT금액']
                            for i in range(1, 11): numeric_intended.append(f'OT{i}')
                            
                            if col not in numeric_intended:
                                # Ensure all strings are stripped and nan-free
                                self.daily_usage_df[col] = self.daily_usage_df[col].astype(str).str.strip().replace(['nan', 'None', 'NULL', '-0.0', '0.0', 'NaN', 'NAN', 'nan.0'], '')
                        
                        # Add/Fix columns for auto-calculation
                        if '검사방법' not in self.daily_usage_df.columns and '검사량' in self.daily_usage_df.columns:
                            # Migrate old string '검사량' (PAUT, UT...) to '검사방법'
                            self.daily_usage_df['검사방법'] = self.daily_usage_df['검사량'].astype(str)
                            self.daily_usage_df['검사량'] = 0.0
                        
                        for col in ['검사량', '단가', '출장비', '일식', '검사비', '수량']:
                            if col not in self.daily_usage_df.columns:
                                self.daily_usage_df[col] = 0.0
                            else:
                                # [ROBUST] Handle strings with commas before conversion
                                def clean_num(s):
                                    if pd.isna(s): return 0.0
                                    try:
                                        # Remove commas and handle currency/unit markers if any
                                        clean_s = str(s).replace(',', '').replace('원', '').strip()
                                        return float(clean_s) if clean_s else 0.0
                                    except: return 0.0
                                    
                                self.daily_usage_df[col] = self.daily_usage_df[col].apply(clean_num)
                            
                        # [SELF-HEALING] Recover historical Inspection Fees that were saved as 0 due to previous bugs
                        # Inspection Fee = (Amount * Unit Price) + Travel Expense + Meal Cost
                        if not self.daily_usage_df.empty:
                            zero_fee_mask = (self.daily_usage_df['검사비'] == 0) | (self.daily_usage_df['검사비'].isna())
                            if zero_fee_mask.any():
                                # Only recover if we have the inputs
                                can_recover_mask = zero_fee_mask & (self.daily_usage_df['검사량'] > 0) & (self.daily_usage_df['단가'] > 0)
                                if can_recover_mask.any():
                                    recovered_fees = (
                                        (self.daily_usage_df.loc[can_recover_mask, '검사량'] * self.daily_usage_df.loc[can_recover_mask, '단가']) + 
                                        self.daily_usage_df.loc[can_recover_mask, '출장비'].fillna(0)
                                    )
                                    self.daily_usage_df.loc[can_recover_mask, '검사비'] = recovered_fees
                                    print(f"DEBUG: Auto-recovered {can_recover_mask.sum()} historical inspection fees.")
                        
                        # [NEW] Compatibility: If '수량' exists but '검사량' is empty/missing, map '수량' to '검사량'
                        if '수량' in self.daily_usage_df.columns:
                             # If both exist, we prefer '수량' if '검사량' is all zeros or NaNs
                             if '검사량' in self.daily_usage_df.columns:
                                 mask = (self.daily_usage_df['검사량'] == 0) | (self.daily_usage_df['검사량'].isna())
                                 self.daily_usage_df.loc[mask, '검사량'] = self.daily_usage_df.loc[mask, '수량']
                             else:
                                 self.daily_usage_df['검사량'] = self.daily_usage_df['수량']
                        
                        # Ensure MaterialID is numeric
                        if 'MaterialID' in self.daily_usage_df.columns:
                            self.daily_usage_df['MaterialID'] = pd.to_numeric(self.daily_usage_df['MaterialID'], errors='coerce')
                        rtk_columns = ['RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크', 
                                      'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타']
                        for col in rtk_columns:
                            if col not in self.daily_usage_df.columns:
                                self.daily_usage_df[col] = 0.0
                            else:
                                # [CRITICAL] Ensure numeric type for smart hiding aggregation
                                self.daily_usage_df[col] = pd.to_numeric(self.daily_usage_df[col], errors='coerce').fillna(0.0)
                        
                        # Also ensure RTK Category column is removed
                        if 'RTK Category' in self.daily_usage_df.columns:
                            self.daily_usage_df = self.daily_usage_df.drop('RTK Category', axis=1)
                        
                except Exception as e:
                    print(f"DEBUG: Failed to load DailyUsage: {e}")
                    self.daily_usage_df = pd.DataFrame(columns=['Date', 'Site', 'MaterialID', 'Usage', 'Note', 'EntryTime',
                                                        'RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크',
                                                        'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타', 'User', '장비명', '검사량',
                                                        '단가', '출장비', '일식', '검사비',
                                                        '차량번호', '주행거리', '차량점검', '차량비고'])
                
                # 5. Migrate vehicle data from daily_usage_df to transactions_df
                if hasattr(self, 'daily_usage_df') and hasattr(self, 'transactions_df'):
                    if not self.daily_usage_df.empty and not self.transactions_df.empty:
                        # Group daily usage by date, site, material to match transactions
                        for _, daily_row in self.daily_usage_df.iterrows():
                            # Handle NaT values safely
                            daily_date = pd.to_datetime(daily_row['Date'], errors='coerce')
                            if pd.isna(daily_date):
                                continue  # Skip rows with invalid dates
                            daily_date = daily_date.normalize()
                            daily_site = str(daily_row['Site']).strip()
                            daily_mat = daily_row['MaterialID']
                            
                            # Find matching transactions
                            mask = (
                                (pd.to_datetime(self.transactions_df['Date'], errors='coerce').dt.normalize() == daily_date) &
                                (self.transactions_df['Site'].str.strip() == daily_site) &
                                (self.transactions_df['MaterialID'] == daily_mat)
                            )
                            
                            # Update vehicle info for matching transactions - ensured string conversion to match Transactions_df dtype
                            if mask.any():
                                self.transactions_df.loc[mask, '차량번호'] = str(daily_row.get('차량번호', '')).strip()
                                self.transactions_df.loc[mask, '주행거리'] = str(daily_row.get('주행거리', '')).strip()
                                self.transactions_df.loc[mask, '차량점검'] = str(daily_row.get('차량점검', '')).strip()
                                self.transactions_df.loc[mask, '차량비고'] = str(daily_row.get('차량비고', '')).strip()


                
                # Add SN column if it doesn't exist (for backward compatibility)
                if 'SN' not in self.materials_df.columns:
                    self.materials_df['SN'] = ''
                
                # Migrate old schema if needed
                if 'Equipment Code' in self.materials_df.columns and '회사코드' not in self.materials_df.columns:
                    self.migrate_old_schema()
                
                # Apply SN extraction from Model Name to existing data
                if not self.materials_df.empty:
                    updated = False
                    for idx, row in self.materials_df.iterrows():
                        model = row.get('모델명', '')
                        sn = row.get('SN', '')
                        new_model, new_sn = self.extract_sn_from_model(model, sn)
                        
                        if str(model) != str(new_model) or str(sn) != str(new_sn):
                            self.materials_df.at[idx, '모델명'] = new_model
                            self.materials_df.at[idx, 'SN'] = new_sn
                            updated = True
                    
                    if updated:
                        self.save_data()

                # 6. Budget
                try:
                    self.budget_df = pd.read_excel(self.db_path, sheet_name='Budget',
                                                   dtype={'Site': str, 'Note': str, 'LaborDetail': str, 'MaterialDetail': str, 'ExpenseDetail': str})
                    self.budget_df = normalize_cols(self.budget_df)
                    # Add missing detail columns for backward compatibility
                    # [FIX] Use 0.0 instead of '' for numeric columns to prevent TypeError: Invalid value '' for dtype 'float64'
                    if 'UnitPrice' not in self.budget_df.columns:
                        self.budget_df['UnitPrice'] = 0.0
                    if 'LaborDetail' not in self.budget_df.columns:
                        self.budget_df['LaborDetail'] = '{}'
                    if 'MaterialDetail' not in self.budget_df.columns:
                        self.budget_df['MaterialDetail'] = '{}'
                    if 'ExpenseDetail' not in self.budget_df.columns:
                        self.budget_df['ExpenseDetail'] = '{}'

                except Exception as e:
                    print(f"DEBUG: Budget sheet not found or load failed: {e}")
                    self.budget_df = pd.DataFrame(columns=['Site', 'Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit', 'Note', 'LaborDetail', 'MaterialDetail', 'ExpenseDetail'])
                
                # [NEW] 7. Settings (Excel-based Rate Management)
                try:
                    self.settings_df = pd.read_excel(self.db_path, sheet_name='Settings')
                    self.settings_df = normalize_cols(self.settings_df)
                    print("DEBUG: Loaded Settings sheet.")
                except Exception as e:
                    print(f"DEBUG: Settings sheet missing or load failed: {e}. Initializing defaults.")
                    # Initialize with hardcoded defaults
                    labor_defaults = [
                        ['Labor', r, '', '', s] for r, s in {
                            "이사": 55250000, "부장": 55250000, "차장": 47670000, "과장": 41170000,
                            "대리": 37920000, "계장": 34670000, "주임": 31420000, "기사": 29250000
                        }.items()
                    ]
                    material_defaults = [
                        ['Material', item, spec, unit, price] for item, spec, unit, price in [
                            ("PT 약품", "세척제", "CAN", 1500), ("PT 약품", "침투제", "CAN", 2300),
                            ("PT 약품", "현상제", "CAN", 2000), ("MT 약품", "백색페인트", "CAN", 2350),
                            ("MT 약품", "흑색자분", "CAN", 1800), ("방사선투과검사 필름", "MX125", "매", 990),
                            ("글리세린", "20L", "통", 100000), ("필름 현상액", "3L", "통", 16500),
                            ("필름 정착액", "3L", "통", 16500), ("수적방지액", "200mL", "통", 2500)
                        ]
                    ]
                    expense_defaults = [
                        ['Expense', '차량유지비', '주유, 수리, 통행, 주차 등', '일', 150000 // 30],
                        ['Expense', '소모품비', '장갑,일회용 작업복외', '일', 15000 // 30],
                        ['Expense', '복리후생비', '생수, 음료 외 기타', '일', 50000 // 30],
                        ['Expense', 'Se-175', '방사성동위원소 구매', 'EA', 10000000 // 280]
                    ]
                    outsource_defaults = [
                        ['Outsource', '케이엔디이', '방사선투과검사', '공수', 15000]
                    ]
                    self.settings_df = pd.DataFrame(labor_defaults + material_defaults + expense_defaults + outsource_defaults, 
                                                   columns=['Category', 'Name', 'Spec', 'Unit', 'Rate'])
                    self.save_data() # Save the newly created sheet

                # [NEW] Ensure budget columns are numeric to prevent NaN assignment errors in some pandas versions
                numeric_cols = ['Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit']
                for col in numeric_cols:
                    if col in self.budget_df.columns:
                        self.budget_df[col] = pd.to_numeric(self.budget_df[col], errors='coerce').fillna(0).astype(float)
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.show_error_dialog("Error", f"데이터를 불러오는데 실패했습니다: {e}")
            self.materials_df = pd.DataFrame(columns=[
                'MaterialID', '회사코드', '관리품번', '품목명', 'SN', '창고',
                '모델명', '규격', '품목군코드', '공급업체', '제조사', '제조국', 
                '가격', '원가', '관리단위', '수량', '재고하한', 'Active'
            ])
            self.transactions_df = pd.DataFrame(columns=['Date', 'MaterialID', 'Site', 'Type', 'Quantity', 'Note', 'User', '차량번호', '주행거리', '차량점검', '차량비고'])
            self.monthly_usage_df = pd.DataFrame(columns=['MaterialID', 'Year', 'Month', 'Site', 'Usage', 'Note', 'Entry Date'])
            self.daily_usage_df = pd.DataFrame(columns=['Date', 'Site', 'MaterialID', 'Usage', 'Note', 'EntryTime',
                                                'RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크',
                                                'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타', '장비명', '검사량', '회사코드',
                                                '업체명', '단가', '출장비', '일식', '검사비', 'FilmCount',
                                                '차량번호', '주행거리', '차량점검', '차량비고'])
            self.budget_df = pd.DataFrame(columns=['Site', 'Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit', 'Note', 'LaborDetail', 'MaterialDetail', 'ExpenseDetail'])

        
        # --- [NEW] Global Data Sanitization & Dtype Enforcement ---
        # This definitively prevents "TypeError: Invalid value '' for dtype 'float64'" by ensuring 
        # that all numeric columns are strictly float64 and free of empty strings or text.
        sanitization_map = {
            'budget_df': ['Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit'],
            'daily_usage_df': ['Usage', '검사량', '단가', '출장비', '일식', '검사비', '수량', 
                                'RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크', 
                                'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타'],
            'materials_df': ['가격', '원가', '수량', '재고하한'],
            'transactions_df': ['Quantity']
        }
        
        for attr, cols in sanitization_map.items():
            if hasattr(self, attr):
                df = getattr(self, attr)
                if df is not None:
                    for col in cols:
                        if col in df.columns:
                            # Convert to numeric, turn errors to NaN, then NaN to 0.0, then force float
                            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0).astype(float)
                    setattr(self, attr, df)

        # Global header cleanup: remove ALL internal and edge whitespace for permanent stability
        for df_attr in ['materials_df', 'transactions_df', 'daily_usage_df', 'budget_df']:
            if hasattr(self, df_attr):
                df = getattr(self, df_attr)
                if df is not None and not df.empty:
                    df.columns = [re.sub(r'\s+', '', str(c)) for c in df.columns]
                    setattr(self, df_attr, df)
            
        # Refresh all inquiry filters once data is loaded
        try:
            self.refresh_inquiry_filters()
        except Exception as e:
            print(f"DEBUG: Initial refresh_inquiry_filters failed: {e}")


    
    def migrate_old_schema(self):
        """Migrate data from old schema to new schema"""
        old_df = self.materials_df.copy()
        self.materials_df = pd.DataFrame(columns=[
            'MaterialID', '회사코드', '관리품번', '품목명', 'SN', '창고',
            '모델명', '규격', '품목군코드', '제조사', '제조국', 
            '가격', '원가', '관리단위', '수량'
        ])
        
        # Headers are already cleaned by the global cleanup in load_data/init

        for _, row in old_df.iterrows():
            new_row = {
                'MaterialID': row.get('MaterialID', ''),
                '회사코드': '',
                '관리품번': row.get('Equipment Code', ''),
                '품목명': row.get('Item Name', row.get('Name', '')),
                'SN': row.get('SN', ''),
                '창고': '',
                '모델명': '',
                '규격': row.get('Specification', ''),
                '품목군코드': '',
                '제조사': row.get('Manufacturer', ''),
                '제조국': '',
                '가격': 0,
                '관리단위': row.get('Unit', 'EA'),
                '수량': row.get('Current Stock', row.get('Initial Stock', 0))
            }
            self.materials_df = pd.concat([self.materials_df, pd.DataFrame([new_row])], ignore_index=True)
        
        self.save_data()
        messagebox.showinfo("마이그레이션 완료", "기존 데이터가 새로운 형식으로 변환되었습니다.")

    def register_new_material(self, name, model='', sn='', **kwargs):
        """Helper to register a new material in the master list summerly"""
        try:
            # Robustly calculate NEW ID even if mixed types or NaNs exist
            valid_ids = pd.to_numeric(self.materials_df['MaterialID'], errors='coerce').dropna()
            new_id = int(valid_ids.max() + 1) if not valid_ids.empty else 10001
        except:
            new_id = 10001 # Fallback
        
        # Determine defaults
        new_mat = {
            'MaterialID': new_id,
            '회사코드': kwargs.get('co_code', ''),
            '관리품번': '',
            '품목명': name,
            'SN': sn,
            '창고': kwargs.get('warehouse', ''),
            '모델명': model or name,
            '규격': '',
            '품목군코드': '',
            '공급업체': '',
            '제조사': '',
            '제조국': '',
            '가격': 0,
            '원가': 0,
            '관리단위': 'EA',
            '수량': 0,
            '재고하한': 10,
            'Active': 1
        }
        # Update with any explicit kwargs
        for k, v in kwargs.items():
            if k in new_mat: new_mat[k] = v
            
        self.materials_df = pd.concat([self.materials_df, pd.DataFrame([new_mat])], ignore_index=True)
        self.save_data()
        self.update_material_combo()
        return new_id
        
        self.save_data()
        messagebox.showinfo("마이그레이션 완료", "기존 데이터가 새로운 형식으로 변환되었습니다.")
    
    def enable_autocomplete(self, combobox, values_list_attr=None, values_list=None, prefix_list=None):
        """Standardize to Native Dropdown: Filter values but use standard system dropdown."""
        if not hasattr(self, '_autocomplete_timers'):
            self._autocomplete_timers = {}

        def perform_filter(event_widget, force_all=False):
            """Filter combobox values based on typed text."""
            typed = str(event_widget.get()).strip()
            widget_name = str(event_widget)
            
            if not typed and not force_all:
                return  # Don't filter on empty input unless forced
            
            if values_list_attr:
                base_values = getattr(self, values_list_attr, [])
            elif values_list is not None:
                base_values = values_list
            else:
                base_values = []
            
            all_values = (prefix_list + base_values) if prefix_list else base_values
            
            if force_all or not typed:
                new_values = all_values
            elif prefix_list and typed in prefix_list:
                new_values = all_values
            else:
                new_values = [v for v in all_values if typed.lower() in str(v).lower()]
            
            print(f"[FILTER] new_values count: {len(new_values)}, changed: {list(event_widget['values']) != list(new_values)}")
            
            # Update values if changed
            if list(event_widget['values']) != list(new_values):
                event_widget['values'] = new_values
            
            # Auto-open dropdown after typing (only if not empty and has focus)
            if not force_all and typed and len(new_values) > 0:
                # Cancel any pending open
                if hasattr(event_widget, '_dropdown_timer'):
                    event_widget.after_cancel(event_widget._dropdown_timer)
                
                def open_dropdown():
                    try:
                        # [STABILITY] Check if it was recently selected to avoid "sticky" dropdown
                        if getattr(event_widget, '_just_selected', False):
                            return

                        # Only if still has focus and NOT already open (heuristic)
                        if event_widget.focus_get() == event_widget:
                            pos = event_widget.index(tk.INSERT)
                            event_widget.event_generate('<Alt-Down>')
                            # Keep typing position
                            event_widget.after(50, lambda: event_widget.icursor(pos))
                    except Exception:
                        pass
                
                event_widget._dropdown_timer = event_widget.after(400, open_dropdown)


        def on_keyrelease(event):
            """Filter values as user types. Down arrow or click ▼ to see filtered list."""
            # Cancel any pending dropdown open while typing
            if hasattr(combobox, '_dropdown_timer'):
                combobox.after_cancel(combobox._dropdown_timer)
                delattr(combobox, '_dropdown_timer')
            
            if event.keysym in ("Left", "Right", "Up", "Down", "Return", "Escape", "Tab", "Shift_L", "Shift_R", "Control_L", "Control_R"):
                return
            
            # Use a timer to wait for typing to stop before filtering
            if combobox in self._autocomplete_timers:
                self.root.after_cancel(self._autocomplete_timers[combobox])
            
            def delayed_filter():
                try:
                    # Poll for value to handle Korean IME composition
                    max_polls = 20
                    poll_count = [0]
                    last_val = ['']
                    
                    def poll_for_value():
                        try:
                            current_val = combobox.get()
                            if not last_val[0] and current_val:
                                perform_filter(combobox, force_all=False)
                                return
                            if not current_val and poll_count[0] < max_polls:
                                last_val[0] = current_val
                                poll_count[0] += 1
                                combobox.after(50, poll_for_value)
                                return
                            if current_val:
                                perform_filter(combobox, force_all=False)
                            else:
                                perform_filter(combobox, force_all=True)
                        except Exception: pass
                    
                    poll_for_value()
                except Exception: pass
                
            self._autocomplete_timers[combobox] = self.root.after(200, delayed_filter)

        combobox.bind('<KeyRelease>', on_keyrelease, add='+')

        # [NEW] Guard against "sticky" dropdown after selection
        def _on_selected(e=None):
            combobox._just_selected = True
            def _reset(): combobox._just_selected = False
            combobox.after(500, _reset)
            
        combobox.bind('<<ComboboxSelected>>', _on_selected, add='+')
        
        # [NEW] Also bind to KeyPress for Korean input composition support
        def on_keypress(event):
            """Handle KeyPress for Korean input composition."""
            print(f"[KEYPRESS] keysym={event.keysym}, char='{event.char}'")
            # For Korean input, schedule filter on any keypress too
            if event.keysym not in ("Left", "Right", "Up", "Down", "Return", "Escape", "Tab", "Shift_L", "Shift_R", "Control_L", "Control_R"):
                # Cancel existing timer
                if combobox in self._autocomplete_timers:
                    self.root.after_cancel(self._autocomplete_timers[combobox])
                
                # Also cancel dropdown timer
                if hasattr(combobox, '_dropdown_timer'):
                    combobox.after_cancel(combobox._dropdown_timer)
                    delattr(combobox, '_dropdown_timer')
                
                # Schedule filter with polling for Korean IME composition
                def delayed_filter_korean():
                    try:
                        # Poll for Korean IME composition - check every 50ms for up to 1 second
                        max_polls = 20
                        poll_count = [0]
                        last_val = ['']
                        
                        def poll_for_value():
                            try:
                                current_val = combobox.get()
                                print(f"[KOREAN-POLL] count={poll_count[0]}, value='{current_val}'")
                                
                                # If value appeared, filter immediately
                                if not last_val[0] and current_val:
                                    print(f"[KOREAN-POLL] Value appeared: '{current_val}'")
                                    perform_filter(combobox, force_all=False)
                                    return
                                
                                # Keep polling while empty
                                if not current_val and poll_count[0] < max_polls:
                                    last_val[0] = current_val
                                    poll_count[0] += 1
                                    combobox.after(50, poll_for_value)
                                    return
                                
                                # Done polling
                                if current_val:
                                    perform_filter(combobox, force_all=False)
                                else:
                                    # [KOREAN IME] Show all values
                                    print(f"[KOREAN-POLL] Timeout, showing all values")
                                    perform_filter(combobox, force_all=True)
                            except Exception as e:
                                print(f"[KOREAN-POLL] Error: {e}")
                        
                        poll_for_value()
                    except Exception as e:
                        print(f"[KOREAN] Error: {e}")
                
                timer_id = self.root.after(400, delayed_filter_korean)  # 400ms for Korean IME
                self._autocomplete_timers[combobox] = timer_id
        
        combobox.bind('<KeyPress>', on_keypress, add=True)

        def on_focus_in(event):
            try:
                combobox.selection_range(0, tk.END)
            except: pass
            perform_filter(combobox, force_all=True)
            
        combobox.bind('<FocusIn>', on_focus_in, add=True)
        
        # Click to show full list
        combobox.bind('<Button-1>', lambda e: perform_filter(combobox, force_all=True), add=True)


    def apply_autocomplete_to_all_comboboxes(self):
        """Map specific comboboxes to their data lists for autocomplete"""
        # Mapping: {widget_attr_name: values_list_attr_name}
        mappings = {
            'cb_material': 'materials_display_list',
            'cb_trans_filter_mat': 'materials_display_list',
            'cb_trans_filter_site': 'sites',
            'cb_budget_site': 'budget_sites',
            'cb_daily_filter_worker': 'users',
            'cb_daily_filter_vehicle': 'vehicles',
            'cb_daily_test_method': 'test_methods',  # Test method list
            'cb_daily_co_code': 'co_code_list',  # [NEW] Company code for MT/PT
            'cb_trans_filter_vehicle': 'vehicles',
            'cb_sales_filter_site': 'sites',
            'cb_filter_co': 'co_code_list',
            'cb_filter_class': 'class_list',
            'cb_filter_mfr': 'mfr_list',
            'cb_filter_name': 'consumable_display_list', # [NEW] Use consumable-only list for Stock View
            'cb_filter_sn': 'sn_list',
            'cb_filter_model': 'model_list',
            'cb_filter_eq': 'eq_code_list',
            'cb_co_code': 'co_code_list',
            'cb_eq_code': 'eq_code_list',
            'cb_item_name': 'item_name_list',
            'cb_model': 'model_list',
            'cb_class': 'class_list',
            'cb_spec': 'spec_list',
            'cb_unit': 'unit_list',
            'cb_supplier': 'supplier_list',
            'cb_mfr': 'mfr_list',
            'cb_origin': 'origin_list',
            'cb_trans_site': 'sites',
            'cb_warehouse': 'warehouses',
            'ent_user': 'users',
            # [NEW] Additional comboboxes for autocomplete
            'cb_name': 'users',  # Worker name
            'ent_worktime': 'worktimes',  # Work time
            'cb_vehicle_info': 'vehicles',  # Vehicle info
        }
        
        for widget_attr, list_attr in mappings.items():
            if hasattr(self, widget_attr):
                widget = getattr(self, widget_attr)
                if isinstance(widget, ttk.Combobox):
                    # For filtering, we might want to include '전체' dynamically
                    current_values = list(widget['values'])
                    if '전체' in current_values:
                        self.enable_autocomplete(widget, values_list_attr=list_attr, prefix_list=['전체'])
                    else:
                        self.enable_autocomplete(widget, values_list_attr=list_attr)

    def _safe_format_datetime(self, val, format_str='%Y-%m-%d %H:%M'):
        if pd.isna(val) or val is None or str(val).strip().lower() in ['nan', 'none', '']:
            return ""
        try:
            return pd.to_datetime(val).strftime(format_str)
        except:
            return str(val)

    def clean_nan(self, val):
        """Robustly clean NaN, None, and other empty markers from any value"""
        if pd.isna(val) or val is None: return ""
        s = str(val).strip()
        if not s or NAN_PATTERN.match(s):
            return ""
        # Handle trailing .0 from numeric inputs converted to string
        s = DOT_ZERO_PATTERN.sub('', s)
        if NAN_PATTERN.match(s): return ""
        return s

    def normalize_id(self, val):
        """Robustly normalize IDs: handle NaN, trailing .0, and whitespace."""
        if pd.isna(val) or val == '' or str(val).lower() == 'nan': return ""
        s = str(val).strip()
        if s.endswith('.0'): s = s[:-2]
        return s

    def save_data(self):
        try:
            # [STABILITY] Ensure MaterialID is safely normalized (numeric where possible, otherwise original string)
            def normalize_id(val):
                if pd.isna(val) or str(val).strip() == '': return val
                try:
                    s_val = str(val).strip()
                    # Handle "10001.0" or "10001"
                    num = float(s_val)
                    if num == int(num): return int(num)
                    return num
                except:
                    # Keep as string (for PAUT/manual names)
                    return str(val).strip()

            for df_name, df in [('Materials', self.materials_df), ('Transactions', self.transactions_df), 
                                ('Monthly', self.monthly_usage_df), ('Daily', self.daily_usage_df)]:
                if df is not None and 'MaterialID' in df.columns:
                    df['MaterialID'] = df['MaterialID'].apply(normalize_id)
                if df is not None and 'Active' in df.columns:
                    df['Active'] = pd.to_numeric(df['Active'], errors='coerce').fillna(1).astype(int)
            
            # [STABILITY] Robust saving with Retry and Backup-on-Fail logic
            max_retries = 3
            retry_delay = 0.5
            
            save_success = False
            last_err = None
            
            for attempt in range(max_retries):
                try:
                    # Explicitly check for write permission/locks
                    if os.path.exists(self.db_path):
                        with open(self.db_path, 'a'): pass
                    
                    with pd.ExcelWriter(self.db_path, engine='openpyxl') as writer:
                        self.materials_df.to_excel(writer, sheet_name='Materials', index=False)
                        self.transactions_df.to_excel(writer, sheet_name='Transactions', index=False)
                        self.monthly_usage_df.to_excel(writer, sheet_name='MonthlyUsage', index=False)
                        self.daily_usage_df.to_excel(writer, sheet_name='DailyUsage', index=False)
                        self.budget_df.to_excel(writer, sheet_name='Budget', index=False)
                        if hasattr(self, 'settings_df'):
                            self.settings_df.to_excel(writer, sheet_name='Settings', index=False)
                    
                    save_success = True
                    break # Success!
                except PermissionError as pe:
                    last_err = pe
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay) # Wait and try again
                        continue
                    else:
                        # Final attempt failed due to lock. Try saving as Conflict Backup.
                        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                        conflict_path = self.db_path.replace('.xlsx', f'_Conflict_{ts}.xlsx')
                        try:
                            with pd.ExcelWriter(conflict_path, engine='openpyxl') as writer:
                                self.materials_df.to_excel(writer, sheet_name='Materials', index=False)
                                self.transactions_df.to_excel(writer, sheet_name='Transactions', index=False)
                                self.monthly_usage_df.to_excel(writer, sheet_name='MonthlyUsage', index=False)
                                self.daily_usage_df.to_excel(writer, sheet_name='DailyUsage', index=False)
                                self.budget_df.to_excel(writer, sheet_name='Budget', index=False)
                                if hasattr(self, 'settings_df'):
                                    self.settings_df.to_excel(writer, sheet_name='Settings', index=False)
                            
                            self.show_error_dialog("데이터 저장 지연/충돌", 
                                f"원본 파일('{os.path.basename(self.db_path)}')이 다른 프로그램에 의해 잠겨 있습니다.\n\n"
                                f"데이터 유실 방지를 위해 다음 경로에 임시 저장되었습니다:\n{os.path.basename(conflict_path)}\n\n"
                                f"동기화 중 오류일 수 있으니, 나중에 수동으로 이름을 변경하거나 파일을 병합해 주세요.")
                            return True # Technically "saved" somewhere
                        except Exception as e2:
                             raise Exception(f"원본 파일 잠김 및 백업 저장 실패: {e2}") from pe
                except Exception as e:
                    raise e # Other errors

            return save_success
        except Exception as e:
            import traceback
            err_detail = traceback.format_exc()
            self.show_error_dialog("데이터 저장 실패", f"데이터를 저장하는데 실패했습니다:\n{e}\n\n파일이 열려있다면 닫고 다시 시도해주세요.")
            return False
            
    def get_base_salaries(self):
        """Extract labor base salaries from settings_df"""
        if not hasattr(self, 'settings_df') or self.settings_df.empty:
            return {
                "이사": 55250000, "부장": 55250000, "차장": 47670000, "과장": 41170000,
                "대리": 37920000, "계장": 34670000, "주임": 31420000, "기사": 29250000
            }
        df = self.settings_df[self.settings_df['Category'] == 'Labor']
        if df.empty:
            return {
                "이사": 55250000, "부장": 55250000, "차장": 47670000, "과장": 41170000,
                "대리": 37920000, "계장": 34670000, "주임": 31420000, "기사": 29250000
            }
        return df.set_index('Name')['Rate'].to_dict()

    def get_material_defaults(self):
        """Extract material defaults from settings_df"""
        if not hasattr(self, 'settings_df') or self.settings_df.empty:
            return [
                ("PT 약품", "세척제", "CAN", 1500), ("PT 약품", "침투제", "CAN", 2300),
                ("PT 약품", "현상제", "CAN", 2000), ("MT 약품", "백색페인트", "CAN", 2350),
                ("MT 약품", "흑색자분", "CAN", 1800), ("방사선투과검사 필름", "MX125", "매", 990),
                ("글리세린", "20L", "통", 100000), ("필름 현상액", "3L", "통", 16500),
                ("필름 정착액", "3L", "통", 16500), ("수적방지액", "200mL", "통", 2500)
            ]
        df = self.settings_df[self.settings_df['Category'] == 'Material']
        if df.empty:
             return [
                ("PT 약품", "세척제", "CAN", 1500), ("PT 약품", "침투제", "CAN", 2300),
                ("PT 약품", "현상제", "CAN", 2000), ("MT 약품", "백색페인트", "CAN", 2350),
                ("MT 약품", "흑색자분", "CAN", 1800), ("방사선투과검사 필름", "MX125", "매", 990),
                ("글리세린", "20L", "통", 100000), ("필름 현상액", "3L", "통", 16500),
                ("필름 정착액", "3L", "통", 16500), ("수적방지액", "200mL", "통", 2500)
            ]
        return [tuple(x) for x in df[['Name', 'Spec', 'Unit', 'Rate']].values]

    def get_expense_defaults(self):
        """Extract site expense defaults from settings_df"""
        if not hasattr(self, 'settings_df') or self.settings_df.empty:
            return [
                ("차량유지비", "주유, 수리, 통행, 주차 등", "N/A", 1, "일", 5000),
                ("소모품비", "장갑,일회용 작업복외", "N/A", 1, "일", 500),
                ("복리후생비", "생수, 음료 외 기타", "N/A", 1, "일", 1667),
                ("Se-175", "방사성동위원소 구매", "N/A", 1, "EA", 35714)
            ]
        df = self.settings_df[self.settings_df['Category'] == 'Expense']
        if df.empty:
            return [
                ("차량유지비", "주유, 수리, 통행, 주차 등", "N/A", 1, "일", 5000),
                ("소모품비", "장갑,일회용 작업복외", "N/A", 1, "일", 500),
                ("복리후생비", "생수, 음료 외 기타", "N/A", 1, "일", 1667),
                ("Se-175", "방사성동위원소 구매", "N/A", 1, "EA", 35714)
            ]
        # Return in (cat, cont, ppl, qty, unit, price) format as expected by _add_row_s1
        return [tuple([x[0], x[1], "N/A", 1, x[2], x[3]]) for x in df[['Name', 'Spec', 'Unit', 'Rate']].values]

    def get_outsource_defaults(self):
        """Extract outsource defaults from settings_df"""
        if not hasattr(self, 'settings_df') or self.settings_df.empty:
            return [("케이엔디이", "방사선투과검사", 0, 15000)]
        df = self.settings_df[self.settings_df['Category'] == 'Outsource']
        if df.empty:
            return [("케이엔디이", "방사선투과검사", 0, 15000)]
        return [tuple(x) for x in df[['Name', 'Spec', 'Unit', 'Rate']].values]

    def create_widgets(self):
        # Notebook for Tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill='both', padx=10, pady=10)
        
        # Tab 1: Current Stock
        self.tab_stock = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_stock, text='현재 재고 현황')
        self.setup_stock_tab()
        
        # Tab 2: Register/Transaction
        self.tab_inout = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_inout, text='입출고 관리')
        self.setup_inout_tab()
        

        # Tab 4: Import/Export
        self.tab_import = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_import, text='데이터 가져오기/내보내기')
        self.setup_import_tab()
        
        # Tab 5: Monthly Usage Entry
        self.tab_monthly_usage = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_monthly_usage, text='월별 집계')
        self.setup_monthly_usage_tab()
        
        # Tab 6: Daily Usage Entry by Site
        self.tab_daily_usage = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_daily_usage, text='현장별 일일 사용량 기입')
        self.setup_daily_usage_tab()
        

        # Tab 8: Project Execution Budget (New)
        self.tab_budget = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_budget, text='공사실행예산서')
        self.setup_budget_tab()
        
        # Initial view update (Move here to ensure all tabs are ready before refresh)
        self.refresh_inquiry_filters()
        self.update_daily_usage_view()
        
        # Bind tab events (Drag and drop reordering)
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        self.notebook.bind("<Button-1>", self.on_tab_drag_start)
        self.notebook.bind("<B1-Motion>", self.on_tab_drag)
        self.notebook.bind("<ButtonRelease-1>", self.on_tab_drag_end)
        self.notebook.bind("<Button-3>", self.show_tab_context_menu)
        
        # Bind main window resize to maintain sash ratios
        self.root.bind("<Configure>", self._on_main_window_resize)
        
        # Load and restore all configuration (geometry, locks, sashes, draggable items)
        self.root.after(100, self.load_tab_config)
        
        # Apply autocomplete to all comboboxes
        self.root.after(200, self.apply_autocomplete_to_all_comboboxes)
        
        # Save tab config on window close
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    def setup_stock_tab(self):
        # Control Frame (Vertical container for both rows)
        control_frame = ttk.Frame(self.tab_stock)
        control_frame.pack(fill='x', padx=5, pady=5)
        
        # Row 1: Action Buttons
        action_row = ttk.Frame(control_frame)
        action_row.pack(fill='x', side='top', pady=(0, 5))
        
        btn_refresh = ttk.Button(action_row, text="재고 새로고침", command=self.update_stock_view)
        btn_refresh.pack(side='left', padx=5)
        
        btn_alert = ttk.Button(action_row, text="재주문 필요 항목 보기", command=self.show_low_stock)
        btn_alert.pack(side='left', padx=5)
        
        btn_delete = ttk.Button(action_row, text="품목 삭제", command=self.delete_selected_material)
        btn_delete.pack(side='left', padx=5)
        
        btn_edit = ttk.Button(action_row, text="품목 수정", command=self.open_edit_material_dialog)
        btn_edit.pack(side='left', padx=5)
        
        btn_export = ttk.Button(action_row, text="엑셀 내보내기", command=self.export_stock_to_excel)
        btn_export.pack(side='left', padx=5)
        
        btn_select_all = ttk.Button(action_row, text="전체 선택", command=self.select_all_stock)
        btn_select_all.pack(side='left', padx=5)
        
        # Row 2: Search and Filter Frame
        filter_row = ttk.Frame(control_frame)
        filter_row.pack(fill='x', side='top')
        
        filter_frame = ttk.LabelFrame(filter_row, text="검색 필터")
        filter_frame.pack(fill='x', padx=5, pady=2)
        
        # Row 0 of Filter Frame (Grid)
        ttk.Label(filter_frame, text="회사:").grid(row=0, column=0, padx=2, pady=2, sticky='e')
        self.cb_filter_co = ttk.Combobox(filter_frame, width=15)
        self.cb_filter_co.grid(row=0, column=1, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="분류:").grid(row=0, column=2, padx=2, pady=2, sticky='e')
        self.cb_filter_class = ttk.Combobox(filter_frame, width=15)
        self.cb_filter_class.grid(row=0, column=3, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="제조사:").grid(row=0, column=4, padx=2, pady=2, sticky='e')
        self.cb_filter_mfr = ttk.Combobox(filter_frame, width=15)
        self.cb_filter_mfr.grid(row=0, column=5, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="품목명:").grid(row=0, column=6, padx=2, pady=2, sticky='e')
        self.cb_filter_name = ttk.Combobox(filter_frame, width=25)
        self.cb_filter_name.grid(row=0, column=7, padx=2, pady=2)
        
        # Row 1 of Filter Frame
        ttk.Label(filter_frame, text="S/N:").grid(row=1, column=0, padx=2, pady=2, sticky='e')
        self.cb_filter_sn = ttk.Combobox(filter_frame, width=20)
        self.cb_filter_sn.grid(row=1, column=1, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="모델명:").grid(row=1, column=2, padx=2, pady=2, sticky='e')
        self.cb_filter_model = ttk.Combobox(filter_frame, width=20)
        self.cb_filter_model.grid(row=1, column=3, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="관리품번:").grid(row=1, column=4, padx=2, pady=2, sticky='e')
        self.cb_filter_eq = ttk.Combobox(filter_frame, width=20)
        self.cb_filter_eq.grid(row=1, column=5, padx=2, pady=2)

        # [NEW] Bind actions to all stock filter comboboxes for consistency and focus handling
        stock_filters = [
            self.cb_filter_co, self.cb_filter_class, self.cb_filter_mfr,
            self.cb_filter_name, self.cb_filter_sn, self.cb_filter_model,
            self.cb_filter_eq
        ]

        def on_stock_filter_action(e):
            self.update_stock_view()
            # [NEW] Move focus to the next filter in sequence for "sideways" navigation
            try:
                current_idx = stock_filters.index(e.widget)
                if current_idx + 1 < len(stock_filters):
                    stock_filters[current_idx + 1].focus_set()
                else:
                    # Move to the General Search entry after all dropdowns
                    self.search_entry.focus_set()
            except:
                # Fallback to result list if sequence fails
                self.stock_tree.focus_set()

        for combo in stock_filters:
            combo.bind("<Return>", on_stock_filter_action)
            combo.bind('<<ComboboxSelected>>', lambda e: self.update_stock_view())
        
        ttk.Label(filter_frame, text="검색어:").grid(row=1, column=6, padx=2, pady=2, sticky='e')
        self.search_var = tk.StringVar()
        self.search_var.trace_add('write', lambda *args: self.update_stock_view())
        self.search_entry = ttk.Entry(filter_frame, textvariable=self.search_var, width=20)
        self.search_entry.grid(row=1, column=7, padx=2, pady=2)
        
        # [NEW] Special handler for Search Entry to move to Results List
        def on_search_enter(e):
            self.update_stock_view()
            self.stock_tree.focus_set()
            if self.stock_tree.get_children():
                # Highlight first item for immediate keyboard control
                self.stock_tree.selection_set(self.stock_tree.get_children()[0])
        self.search_entry.bind("<Return>", on_search_enter)
        
        # Reset Filters Button
        btn_reset = ttk.Button(filter_frame, text="♻️ 필터 초기화", command=self.reset_stock_filters)
        btn_reset.grid(row=1, column=8, padx=10, pady=2)
        
        # Treeview for Stock with Scrollbars
        tree_frame = ttk.Frame(self.tab_stock)
        tree_frame.pack(expand=True, fill='both', padx=5, pady=5)
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        columns = ('ID', '회사코드', '관리품번', '품목명', 'SN', '창고', '모델명', '규격', '품목군코드', '공급업체', '제조사', '제조국', '가격', '원가', '관리단위', '수량', '재고하한', '상태/위치')
        self.stock_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', 
                                      yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        vsb.config(command=self.stock_tree.yview)
        hsb.config(command=self.stock_tree.xview)
        
        # Column configuration
        col_widths = [60, 80, 100, 180, 90, 90, 120, 120, 90, 120, 120, 80, 80, 80, 80, 80, 80, 130]
        for col, width in zip(columns, col_widths):
            self.stock_tree.heading(col, text=col, command=lambda _col=col: self.treeview_sort_column(self.stock_tree, _col, False))
            # Change stretch=True to stretch=False to allow fixed user-defined widths
            self.stock_tree.column(col, width=width, minwidth=50, stretch=False, anchor='center')
        
        # Bind double-click
        self.stock_tree.bind('<Double-1>', lambda e: self.open_edit_material_dialog())
        
        # Auto-save column widths when user resizes columns
        self.stock_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())
        self.enable_tree_column_drag(self.stock_tree, context_menu_handler=lambda e: self._show_generic_tree_heading_context_menu(e, self.stock_tree))
        
        # Grid layout
        self.stock_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Initial view update will be triggered after update_registration_combos 
        # to ensure filters are properly initialized to "전체" first.
        self.refresh_ui_for_list_change('daily_units')
    
    def show_low_stock(self):
        """Show items with low stock (less than their specific reorder point)"""
        low_stock_items = []
        for _, mat in self.materials_df.iterrows():
            if mat.get('Active', 1) == 0:
                continue
                
            current = self.calculate_current_stock(mat['MaterialID'])
            reorder_point = mat.get('재고하한', 10)
            if pd.isna(reorder_point) or reorder_point <= 0:
                reorder_point = 10 # Default fallback
                
            if pd.notna(current) and current < reorder_point:
                low_stock_items.append((mat.get('품목명', ''), current, reorder_point))
        
        if not low_stock_items:
            messagebox.showinfo("재고 알림", "수량이 10개 미만인 항목이 없습니다.")
        else:
            msg = "다음 항목들의 재고가 부족합니다:\n\n"
            for item, current, reorder in low_stock_items:
                msg += f"• {item}: 현재 {current:g} (필요 수준: {reorder:g})\n"
            messagebox.showwarning("재고 부족", msg)
    
    def select_all_stock(self):
        """Select all items in the stock treeview"""
        all_items = self.stock_tree.get_children()
        self.stock_tree.selection_set(all_items)

    def delete_selected_material(self):
        """선택된 자재 항목들을 재고에서 삭제"""
        selected_items = self.stock_tree.selection()
        
        if not selected_items:
            messagebox.showwarning("선택 오류", "삭제할 품목을 선택해주세요.")
            return
        
        # Confirm deletion
        confirm = messagebox.askyesno("삭제 확인", f"선택한 {len(selected_items)}개의 품목을 재고에서 영구히 삭제하시겠습니까?\n이 작업은 되돌릴 수 없습니다.")
        
        if not confirm:
            return
            
        # Get MaterialIDs to delete
        mat_ids_to_remove = []
        for item in selected_items:
            values = self.stock_tree.item(item, 'values')
            if values:
                # Ensure we match the type of MaterialID in the dataframe
                mat_ids_to_remove.append(type(self.materials_df['MaterialID'].iloc[0])(values[0]))
        
        # Soft delete: Set Active=0 instead of removing from materials_df
        if 'Active' not in self.materials_df.columns:
            self.materials_df['Active'] = 1
            
        # Standardize Active column to numeric to avoid pandas TypeError
        self.materials_df['Active'] = pd.to_numeric(self.materials_df['Active'], errors='coerce').fillna(1)
        
        initial_count = len(self.materials_df[self.materials_df['Active'] != 0])
        
        # Apply deletion (set to 0)
        mask = self.materials_df['MaterialID'].isin(mat_ids_to_remove)
        self.materials_df.loc[mask, 'Active'] = 0
        
        final_count = len(self.materials_df[self.materials_df['Active'] != 0])
        removed_count = initial_count - final_count
        
        if removed_count > 0:
            # Save data and update views
            self.save_data()
            self.update_stock_view()
            self.update_material_combo()
            
            # Optional: Clear transactions related to these materials?
            # For now, let's keep them for history unless explicitly asked.
            
            messagebox.showinfo("완료", f"{removed_count}개의 품목이 삭제되었습니다.")
        else:
            messagebox.showwarning("실패", "데이터프레임에서 항목을 삭제하지 못했습니다.")

    def reset_stock_filters(self):
        """Reset all stock filters to default"""
        self.cb_filter_co.set("전체")
        self.cb_filter_class.set("전체")
        self.cb_filter_mfr.set("전체")
        self.cb_filter_name.set("전체")
        self.cb_filter_sn.set("전체")
        self.cb_filter_model.set("전체")
        self.cb_filter_eq.set("전체")
        self.search_var.set("")
        self.update_stock_view()

    def open_edit_material_dialog(self):
        """Open a dialog to edit the selected material"""
        selection = self.stock_tree.selection()
        if not selection:
            messagebox.showwarning("선택 오류", "수정할 자재를 선택해주세요.")
            return
            
        item = self.stock_tree.item(selection[0])
        mat_values = item['values']
        mat_id = mat_values[0]
        
        # Get full material data from DF (Robust matching)
        target_id_str = self.normalize_id(mat_id)
        mask = self.materials_df['MaterialID'].apply(self.normalize_id) == target_id_str
        matches = self.materials_df[mask]
        
        if matches.empty:
            messagebox.showerror("오류", f"자재 ID '{mat_id}'를 데이터베이스에서 찾을 수 없습니다.")
            edit_win.destroy()
            return
            
        mat_data = matches.iloc[0]
        
        # Create Edit Dialog
        edit_win = tk.Toplevel(self.root)
        edit_win.title("자재 정보 수정")
        edit_win.geometry("550x750")
        edit_win.transient(self.root)
        edit_win.grab_set()
        
        # [NEW] Scrollable area for many fields (User request: button visibility)
        outer_frame = ttk.Frame(edit_win)
        outer_frame.pack(fill='both', expand=True)
        
        canvas = tk.Canvas(outer_frame, highlightthickness=0, bg=getattr(self, 'theme_bg', '#f0f0f0'))
        scrollbar = ttk.Scrollbar(outer_frame, orient="vertical", command=canvas.yview)
        main_frame = ttk.Frame(canvas, padding=20)
        
        main_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=main_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Enable mousewheel scrolling (Safer local binding)
        def _on_mousewheel(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except: pass
        edit_win.bind("<MouseWheel>", _on_mousewheel)
        
        fields = [
            ('회사코드', '회사코드'),
            ('관리품번', '관리품번'),
            ('품목명', '품목명'),
            ('SN', 'SN'),
            ('창고', '창고'),
            ('모델명', '모델명'),
            ('규격', '규격'),
            ('품목군코드', '품목군코드'),
            ('공급업체', '공급업체'),
            ('제조사', '제조사'),
            ('제조국', '제조국'),
            ('가격', '가격'),
            ('원가', '원가'),
            ('관리단위', '관리단위'),
            ('수량', '수량'),
            ('재고하한', '재고하한')
        ]
        
        entries = {}
        for i, (label_text, col_name) in enumerate(fields):
            ttk.Label(main_frame, text=f"{label_text}:").grid(row=i, column=0, padx=5, pady=5, sticky='w')
            
            # Using Combobox for some fields to maintain consistency
            if col_name in ['회사코드', '품목군코드', '제조사', '제조국', '관리단위']:
                ent = ttk.Combobox(main_frame, width=35)
                # [UNIFIED_UNIT] Populate values from managed list for units, others from data
                if col_name == '관리단위':
                    ent['values'] = getattr(self, 'unit_list', self.daily_units)
                elif col_name in self.materials_df.columns:
                    unique_vals = sorted(self.materials_df[col_name].dropna().unique().tolist())
                    ent['values'] = unique_vals
            else:
                ent = ttk.Entry(main_frame, width=38)
                
            ent.grid(row=i, column=1, padx=5, pady=5)
            
            # Pre-fill value
            val = mat_data.get(col_name, '')
            if pd.isna(val): val = ''
            ent.insert(0, str(val))
            entries[col_name] = ent
            
        def on_save():
            new_data = {col: ent.get() for col, ent in entries.items()}
            self.save_material_edits(mat_id, new_data)
            edit_win.destroy()
            
        btn_save = ttk.Button(main_frame, text="변경사항 저장", command=on_save)
        btn_save.grid(row=len(fields), column=0, columnspan=2, pady=20)

    def save_material_edits(self, mat_id, new_data):
        """Save edited material data back to the database"""
        # Update DataFrame (Robust matching)
        target_id_str = self.normalize_id(mat_id)
        mask = self.materials_df['MaterialID'].apply(self.normalize_id) == target_id_str
        idx = self.materials_df.index[mask].tolist()
        if not idx:
            messagebox.showerror("오류", "자재를 찾을 수 없습니다.")
            return
            
        for col, val in new_data.items():
            if col in self.materials_df.columns:
                # Check the actual dtype of the column
                col_dtype = self.materials_df[col].dtype
                
                # Handle different data types appropriately
                if col_dtype == 'float64' or col_dtype == 'int64':
                    # Numeric columns - convert to float, empty becomes 0.0
                    try:
                        val = float(val) if str(val).strip() else 0.0
                    except (ValueError, TypeError):
                        val = 0.0
                else:
                    # String/object columns - handle empty values
                    if val == '' or val == 'nan' or pd.isna(val):
                        val = ''
                    else:
                        val = str(val)
                
                self.materials_df.at[idx[0], col] = val
        
        # Re-check SN extraction after edit
        model = self.materials_df.at[idx[0], '모델명']
        sn = self.materials_df.at[idx[0], 'SN']
        new_model, new_sn = self.extract_sn_from_model(model, sn)
        self.materials_df.at[idx[0], '모델명'] = new_model
        self.materials_df.at[idx[0], 'SN'] = new_sn
        
        # Save to Excel
        self.save_data()
        
        # Refresh everything
        self.update_stock_view()
        self.update_registration_combos()
        self.update_material_combo()
        
        messagebox.showinfo("완료", "자재 정보가 성공적으로 수정되었습니다.")

    
    def calculate_current_stock(self, mat_id):
        """Calculate current stock for a material based on transactions (fully robust version)"""
        # 1. Normalize the target mat_id to clean string
        str_mat_id = re.sub(r'\.0$', '', str(mat_id).strip())
        if not str_mat_id or str_mat_id.lower() == 'nan':
            return 0.0

        # Helper to get normalized MaterialID series
        def get_norm_series(df):
            return df['MaterialID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)

        # 2. Get Transaction Sum
        net_trans_qty = 0.0
        if not self.transactions_df.empty:
            mat_trans = self.transactions_df[get_norm_series(self.transactions_df) == str_mat_id]
            net_trans_qty = float(mat_trans['Quantity'].sum()) if not mat_trans.empty else 0.0
        
        # 3. Get Base Quantity and Model Name from Materials Master
        stored_qty = 0.0
        is_chemical_group = False
        if not self.materials_df.empty:
            mat_rows = self.materials_df[get_norm_series(self.materials_df) == str_mat_id]
            if not mat_rows.empty:
                mat = mat_rows.iloc[0]
                model_name = str(mat.get('모델명', '')).strip().upper()
                if model_name == 'MT약품' and str(mat.get('관리단위', '')).strip().upper() != 'EA':
                    # Only exclude non-consumables (usually EA units are for equipment)
                    # If it's a chemical but mislabeled as MT약품, we might still want to track it if it has a non-EA unit.
                    # However, safer for now is to just fix normalization first. 
                    # Let's keep the user's specific request but fix the ID matching below.
                    return 0.0
                
                val = mat.get('수량', 0)
                try: stored_qty = float(str(val).replace(',', '')) if pd.notna(val) else 0.0
                except: stored_qty = 0.0
                
        return stored_qty + net_trans_qty

    def update_stock_view(self):
        # Clear current view
        for item in self.stock_tree.get_children():
            self.stock_tree.delete(item)
        
        search_term = self.search_var.get().lower() if hasattr(self, 'search_var') else ''
        
        # Helper function to safely get value and replace NaN
        def safe_get(val, default=''):
            cleaned = self.clean_nan(val)
            if cleaned == '' and str(val).lower() in ['nan', 'none', 'null', 'nan.0', '0.0', '-0.0', '']:
                return default
            return val if pd.notna(val) else default
        
        def to_f(val):
            if pd.isna(val) or val is None: return 0.0
            try:
                s = str(val).replace(',', '').strip()
                if not s: return 0.0
                return float(s)
            except:
                return 0.0
        
        # [OPTIMIZATION] Pre-calculate stock lookup to avoid O(N*M) performance hit
        stock_lookup = {}
        if not self.transactions_df.empty:
            temp_trans = self.transactions_df.copy()
            # Use standardized normalization
            temp_trans['NormID'] = temp_trans['MaterialID'].apply(self.normalize_id)
            # [CRITICAL] Exclude "자동 차감" (Automatic Deduction) to avoid double-counting with Daily Usage sheet
            mask = ~temp_trans['Note'].astype(str).str.contains('자동 차감', na=False)
            # Group by normalized ID and sum quantities
            stock_lookup = temp_trans[mask].groupby('NormID')['Quantity'].sum().to_dict()

        # [NEW] Pre-calculate Daily Usage subtraction
        daily_usage_lookup = {}
        daily_name_lookup = {}
        
        # [NEW] Get a set of consumable MaterialIDs from master for robust deduction
        consumable_ids = set()
        if not self.materials_df.empty:
            for _, m in self.materials_df.iterrows():
                # Check if the master item itself is a consumable
                m_name = str(m.get('품목명', '')).strip()
                if self._is_consumable_material(m_name, ''):
                    c_id = self.normalize_id(m.get('MaterialID'))
                    if c_id: consumable_ids.add(c_id)
        
        if hasattr(self, 'daily_usage_df') and not self.daily_usage_df.empty:
            temp_daily = self.daily_usage_df.copy()
            temp_daily['NormID'] = temp_daily['MaterialID'].apply(self.normalize_id)
            
            def _f(v):
                if pd.isna(v) or v is None: return 0.0
                try: return float(str(v).replace(',', '').strip()) if str(v).strip() else 0.0
                except: return 0.0
                
            # [ROBUST] Support multiple column names for usage and film counts
            temp_daily['TotalUsage'] = temp_daily.apply(lambda r: 
                _f(r.get('Usage', r.get('검사량', r.get('수량', r.get('Quantity', 0))))) + 
                _f(r.get('FilmCount', r.get('매수', 0))), axis=1)
            
            # [NEW] Filter: Only include consumables in the site-usage deduction lookup
            # This ensures durable equipment (PAUT, MT-Yoke) reported in the site tab doesn't decrease stock
            # [ROBUST] Use MaterialID as primary signal, fallback to 품목명/장비명 keywords
            temp_daily['IsConsumable'] = temp_daily.apply(lambda r: 
                (r['NormID'] in consumable_ids) or 
                self._is_consumable_material(
                    str(r.get('품목명', r.get('장비명', ''))).strip(), 
                    str(r.get('검사방법', '')).strip()
                ), axis=1)
            
            temp_consumable = temp_daily[temp_daily['IsConsumable']]
            
            daily_usage_lookup = temp_consumable.groupby('NormID')['TotalUsage'].sum().to_dict()
            
            # [NEW] Also build a name-based lookup for fallback
            temp_consumable['NormName'] = temp_consumable.apply(lambda r: str(r.get('품목명', r.get('장비명', ''))).strip(), axis=1)
            daily_name_lookup = temp_consumable.groupby('NormName')['TotalUsage'].sum().to_dict()
            print(f"DEBUG: Daily Consumable Lookup built: {daily_name_lookup}")
            
            # [NEW] Pre-calculate NDT chemical totals by name (Enhanced with fuzzy matching and JSON parsing)
            ndt_name_lookup = {}
            ndt_keys = ['세척제', '침투제', '현상제', '백색페인트', '흑색자분', '형광자분', '형광침투제', '자분페인트']
            for k in ndt_keys: ndt_name_lookup[k] = 0.0

            for _, row in temp_daily.iterrows():
                # 1. Check direct columns (fuzzy match)
                for col in temp_daily.columns:
                    for k in ndt_keys:
                        if k in str(col):
                            ndt_name_lookup[k] += _f(row.get(col, 0))
                
                # 2. Check ndt_data JSON field
                nj_raw = row.get('ndt_data', row.get('ndtdata', ''))
                if nj_raw and isinstance(nj_raw, str) and nj_raw.strip().startswith('{'):
                    try:
                        import json
                        ndt_json = json.loads(nj_raw)
                        for k in ndt_keys:
                            # Search in JSON keys
                            for jk, jv in ndt_json.items():
                                if k in str(jk):
                                    ndt_name_lookup[k] += _f(jv)
                    except: pass
                elif isinstance(nj_raw, dict):
                    for k in ndt_keys:
                        for jk, jv in nj_raw.items():
                            if k in str(jk):
                                ndt_name_lookup[k] += _f(jv)
            
            # Filter out zero values for cleaner lookup
            ndt_name_lookup = {k: v for k, v in ndt_name_lookup.items() if v > 0}
            print(f"DEBUG: Daily Usage lookup built. NDT totals: {ndt_name_lookup}")

        # Calculate current stock
        stock_summary = []
        
        # Pre-setup tags for equipment status highlighting
        self.stock_tree.tag_configure('deployed', background='#FFF9C4') # Light Yellow for "Field"
        self.stock_tree.tag_configure('in_stock', background='') # Default
        
        for _, mat in self.materials_df.iterrows():
            if mat.get('Active', 1) == 0:
                continue
            
            mat_id = mat['MaterialID']
            # [NEW] Format MaterialID as integer for display at the start of loop
            display_id = ""
            try:
                if pd.notna(mat_id) and str(mat_id).strip():
                    display_id = str(int(float(str(mat_id).strip())))
            except:
                display_id = str(mat_id)

            # [REFINED] Skip non-consumables OR auto-registered items that are NOT consumables.
            # We ALLOW auto-registered items if they are confirmed as consumables (like films/drugs).
            spec = str(mat.get('규격', '')).strip()
            mat_name_str = str(mat.get('품목명', mat.get('ǰ', ''))).strip()
            is_consumable = self._is_consumable_material(mat_name_str, '')
            
            if not is_consumable or (spec == "자동등록" and not is_consumable):
                continue
            str_mat_id = self.normalize_id(mat_id)
            
            # Use optimized lookup
            net_trans_qty = stock_lookup.get(str_mat_id, 0.0)
            daily_qty = daily_usage_lookup.get(str_mat_id, 0.0)
            
            mat_name_str = str(mat.get('품목명', mat.get('ǰ', ''))).strip()
            # [NEW] Fallback to name-based lookup if ID lookup is 0
            if daily_qty == 0 and 'daily_name_lookup' in locals():
                daily_qty = daily_name_lookup.get(mat_name_str, 0.0)
            
            # Get stored quantity
            val = mat.get('수량', 0)
            try: stored_qty = float(str(val).replace(',', '')) if pd.notna(val) else 0.0
            except: stored_qty = 0.0
            
            # [FINAL_STOCK_CALC] Current Stock = Master + In/Out - Site Usage
            ndt_usage = 0.0
            mat_name_raw = self.clean_nan(mat.get('품목명', mat.get('ǰ', '')))
            model_name_raw = self.clean_nan(mat.get('모델명', mat.get('𵨸', '')))
            
            # Combine Item Name + Model Name for robust keyword searching (e.g. 'PT약품 세척제')
            combined_name = (mat_name_raw + " " + model_name_raw).replace(' ', '')
            
            # If this is an NDT item, find matching usage by name
            if hasattr(self, 'ndt_name_lookup') or 'ndt_name_lookup' in locals():
                lookup_source = ndt_name_lookup if 'ndt_name_lookup' in locals() else self.ndt_name_lookup
                
                # [REFINED] Only match specific keywords from combined name.
                for ndt_key, ndt_val in lookup_source.items():
                    if ndt_key in combined_name:
                        ndt_usage = ndt_val
                        # If we found a specific NDT chemical (e.g. via model name), 
                        # ensure we don't also subtract generic inspection quantity
                        daily_qty = 0.0 
                        break
            
            current_stock = stored_qty + net_trans_qty - daily_qty - ndt_usage
            
            # Debug for specific items the user is watching (Films and NDT drugs)
            if any(k in combined_name.upper() for k in ['세척제', '침투제', '현상제', '백색', '흑색', '자분', 'CARESTREAM', 'MX125']):
                raw_val = mat.get('수량', 'MISSING')
                print(f"DEBUG: Stock Calc for '{mat_name_raw} ({model_name_raw})': RawQty='{raw_val}', Master={stored_qty}, InOut={net_trans_qty}, Daily={daily_qty}, NDT={ndt_usage}, Final={current_stock}")
            
            # --- Dynamic Location/Status Tracking ---
            status_location = "관내 (창고)"
            row_tag = 'in_stock'
            
            if current_stock <= 0:
                # Look for the last "OUT" transaction to determine where it went
                if not self.transactions_df.empty:
                    # Optimized last transaction check
                    mask = self.transactions_df['MaterialID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True) == str_mat_id
                    relevant_trans = self.transactions_df[mask].sort_values(by='Date', ascending=False)
                    
                    if not relevant_trans.empty:
                        last_op = relevant_trans.iloc[0]
                        if last_op['Type'] == 'OUT':
                            site_name = self.clean_nan(last_op.get('Site', ''))
                            status_location = f"현장: {site_name}" if site_name else "출고됨"
                            row_tag = 'deployed'
            
            # Default row display
            stock_summary.append({
                'data': (
                    display_id,
                    safe_get(mat.get('회사코드', ''), ''),
                    safe_get(mat.get('관리품번', ''), ''),
                    safe_get(mat.get('품목명', ''), ''),
                    safe_get(mat.get('SN', ''), ''),
                    safe_get(mat.get('창고', ''), ''),
                    safe_get(mat.get('모델명', ''), ''),
                    safe_get(mat.get('규격', ''), ''),
                    safe_get(mat.get('품목군코드', ''), ''),
                    safe_get(mat.get('공급업체', ''), ''),
                    safe_get(mat.get('제조사', ''), ''),
                    safe_get(mat.get('제조국', ''), ''),
                    f"{to_f(mat.get('가격', 0)):,.0f}",
                    f"{to_f(mat.get('원가', 0)):,.0f}",
                    safe_get(mat.get('관리단위', 'EA'), 'EA'),
                    f"{to_f(current_stock):g}",
                    f"{to_f(mat.get('재고하한', 0)):g}",
                    status_location
                ),
                'tag': row_tag
            })
        
        # [NEW] Sort by Item Name (Index 3) then Model Name (Index 6)
        stock_summary.sort(key=lambda x: (str(x['data'][3]), str(x['data'][6])))
        
        # Filter by search term and dropdowns
        filter_co = str(self.cb_filter_co.get()).strip() if self.cb_filter_co.get() else "전체"
        filter_class = str(self.cb_filter_class.get()).strip() if self.cb_filter_class.get() else "전체"
        filter_mfr = str(self.cb_filter_mfr.get()).strip() if self.cb_filter_mfr.get() else "전체"
        filter_name = str(self.cb_filter_name.get()).strip() if self.cb_filter_name.get() else "전체"
        filter_sn = str(self.cb_filter_sn.get()).strip() if self.cb_filter_sn.get() else "전체"
        filter_model = str(self.cb_filter_model.get()).strip() if self.cb_filter_model.get() else "전체"
        filter_eq = str(self.cb_filter_eq.get()).strip() if self.cb_filter_eq.get() else "전체"
        
        for row_obj in stock_summary:
            row = row_obj['data']
            # Dropdown Filters
            if filter_co != "전체" and str(row[1]) != filter_co: continue
            if filter_class != "전체" and str(row[8]) != filter_class: continue
            if filter_mfr != "전체" and str(row[10]) != filter_mfr: continue
            if filter_name != "전체" and str(row[3]) != filter_name: continue
            if filter_sn != "전체" and str(row[4]) != filter_sn: continue
            if filter_model != "전체" and str(row[6]) != filter_model: continue
            if filter_eq != "전체" and str(row[2]) != filter_eq: continue
            
            # General Search Term
            if search_term:
                row_str = ' '.join(str(x).lower() for x in row)
                if search_term not in row_str:
                    continue
            self.stock_tree.insert('', tk.END, values=row, tags=(row_obj['tag'],))
        
    def treeview_sort_column(self, tv, col, reverse):
        """Standard sorting function for Treeview columns"""
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        
        # Try to sort numerically if it looks like a number
        try:
            # Clean values like "1,234" or "EA"
            l.sort(key=lambda t: float(str(t[0]).replace(',', '').strip()) if str(t[0]).replace(',', '').strip() else 0, reverse=reverse)
        except (ValueError, TypeError):
            # Fallback to string sort
            l.sort(reverse=reverse)

        # Rearrange items in sorted positions
        for index, (val, k) in enumerate(l):
            tv.move(k, '', index)

        # Switch to reverse sort for next click
        tv.heading(col, command=lambda: self.treeview_sort_column(tv, col, not reverse))
        
    def setup_inout_tab(self):
        # Main PanedWindow (Vertical): Top (Registration) vs Bottom (History)
        self.inout_paned = ttk.Panedwindow(self.tab_inout, orient='vertical')
        self.inout_paned.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Horizontal PanedWindow for Registration Sections (Side-by-Side)
        self.inout_top_paned = ttk.Panedwindow(self.inout_paned, orient='horizontal')
        self.inout_paned.add(self.inout_top_paned, weight=1)
        
        # Save sash position on adjustment
        self.inout_paned.bind("<ButtonRelease-1>", lambda e: self.save_tab_config())
        
        # Side 1: New Material Registration
        reg_container = ttk.Frame(self.inout_top_paned)
        self.inout_top_paned.add(reg_container, weight=1)
        
        # Frame for Registration (Directly packed)
        reg_frame = ttk.LabelFrame(reg_container, text="자재 신규 등록")
        reg_frame.pack(fill='both', expand=True, padx=10, pady=5)
        reg_frame.grid_columnconfigure(1, weight=1)
        reg_frame.grid_columnconfigure(3, weight=1)
        
        # Row 0
        ttk.Label(reg_frame, text="회사코드:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.cb_co_code = ttk.Combobox(reg_frame, width=15)
        self.cb_co_code.grid(row=0, column=1, padx=5, pady=2, sticky='ew')
        self.cb_co_code.bind('<Return>', lambda e: self.cb_eq_code.focus_set())
        
        ttk.Label(reg_frame, text="설비코드:").grid(row=0, column=2, padx=5, pady=2, sticky='w')
        self.cb_eq_code = ttk.Combobox(reg_frame, width=20)
        self.cb_eq_code.grid(row=0, column=3, padx=5, pady=2, sticky='ew')
        self.cb_eq_code.bind('<Return>', lambda e: self.cb_item_name.focus_set())
        
        # Row 1
        ttk.Label(reg_frame, text="품목명:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        item_name_container = ttk.Frame(reg_frame)
        item_name_container.grid(row=1, column=1, padx=5, pady=2, sticky='ew')
        
        self.cb_item_name = ttk.Combobox(item_name_container, width=15)
        self.cb_item_name.pack(side='left', fill='x', expand=True)
        self.cb_item_name.bind('<Return>', lambda e: self.cb_model.focus_set())
        
        # [MODERN] Place the search button INSIDE the combobox
        btn_pref_search = tk.Button(self.cb_item_name, text="🔍", font=('Arial', 8), 
                                    bd=0, bg='white', cursor='hand2',
                                    command=self.open_preferred_item_search_dialog)
        # Offset x=-20 to avoid overlapping the combobox arrow
        btn_pref_search.place(relx=1.0, x=-20, rely=0.5, anchor='e', width=18, height=18)
        
        ttk.Label(reg_frame, text="모델명:").grid(row=1, column=2, padx=5, pady=2, sticky='w')
        self.cb_model = ttk.Combobox(reg_frame, width=20)
        self.cb_model.grid(row=1, column=3, padx=5, pady=2, sticky='ew')
        self.cb_model.bind('<Return>', lambda e: self.ent_sn.focus_set())
        
        # Row 2
        ttk.Label(reg_frame, text="SN번호:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.ent_sn = ttk.Entry(reg_frame, width=15)
        self.ent_sn.grid(row=2, column=1, padx=5, pady=2, sticky='ew')
        self.ent_sn.bind('<Return>', lambda e: self.cb_class.focus_set())
        
        ttk.Label(reg_frame, text="분류:").grid(row=2, column=2, padx=5, pady=2, sticky='w')
        self.cb_class = ttk.Combobox(reg_frame, width=20)
        self.cb_class.grid(row=2, column=3, padx=5, pady=2, sticky='ew')
        self.cb_class.bind('<Return>', lambda e: self.cb_spec.focus_set())
        
        # Row 3
        ttk.Label(reg_frame, text="규격:").grid(row=3, column=0, padx=5, pady=2, sticky='w')
        self.cb_spec = ttk.Combobox(reg_frame, width=15)
        self.cb_spec.grid(row=3, column=1, padx=5, pady=2, sticky='ew')
        self.cb_spec.bind('<Return>', lambda e: self.cb_unit.focus_set())
        
        ttk.Label(reg_frame, text="단위:").grid(row=3, column=2, padx=5, pady=2, sticky='w')
        self.cb_unit = ttk.Combobox(reg_frame, width=20)
        self.cb_unit.grid(row=3, column=3, padx=5, pady=2, sticky='ew')
        self.cb_unit.bind('<Return>', lambda e: self.cb_supplier.focus_set())
        
        # Row 4
        ttk.Label(reg_frame, text="공급업자:").grid(row=4, column=0, padx=5, pady=2, sticky='w')
        self.cb_supplier = ttk.Combobox(reg_frame, width=15)
        self.cb_supplier.grid(row=4, column=1, padx=5, pady=2, sticky='ew')
        self.cb_supplier.bind('<Return>', lambda e: self.cb_mfr.focus_set())
        
        ttk.Label(reg_frame, text="제조사:").grid(row=4, column=2, padx=5, pady=2, sticky='w')
        self.cb_mfr = ttk.Combobox(reg_frame, width=20)
        self.cb_mfr.grid(row=4, column=3, padx=5, pady=2, sticky='ew')
        self.cb_mfr.bind('<Return>', lambda e: self.cb_origin.focus_set())
        
        # Row 5
        ttk.Label(reg_frame, text="제조국:").grid(row=5, column=0, padx=5, pady=2, sticky='w')
        self.cb_origin = ttk.Combobox(reg_frame, width=15)
        self.cb_origin.grid(row=5, column=1, padx=5, pady=2, sticky='ew')
        self.cb_origin.bind('<Return>', lambda e: self.ent_reorder.focus_set())
        
        ttk.Label(reg_frame, text="재주문 수준:").grid(row=5, column=2, padx=5, pady=2, sticky='w')
        self.ent_reorder = ttk.Entry(reg_frame, width=20)
        self.ent_reorder.grid(row=5, column=3, padx=5, pady=2, sticky='ew')
        self.ent_reorder.insert(0, "0")
        self.ent_reorder.bind('<Return>', lambda e: self.ent_init.focus_set())
        
        # Row 6
        ttk.Label(reg_frame, text="초기재고:").grid(row=6, column=0, padx=5, pady=2, sticky='w')
        self.ent_init = ttk.Entry(reg_frame, width=15)
        self.ent_init.grid(row=6, column=1, padx=5, pady=2, sticky='ew')
        self.ent_init.insert(0, "0")
        self.ent_init.bind('<Return>', lambda e: self.ent_price.focus_set())
        
        ttk.Label(reg_frame, text="단가(가격):").grid(row=6, column=2, padx=5, pady=2, sticky='w')
        self.ent_price = ttk.Entry(reg_frame, width=20)
        self.ent_price.grid(row=6, column=3, padx=5, pady=2, sticky='ew')
        self.ent_price.insert(0, "0")
        self.ent_price.bind('<Return>', lambda e: self.ent_cost.focus_set())
        
        # Row 7
        ttk.Label(reg_frame, text="원가:").grid(row=7, column=0, padx=5, pady=2, sticky='w')
        self.ent_cost = ttk.Entry(reg_frame, width=15)
        self.ent_cost.grid(row=7, column=1, padx=5, pady=2, sticky='ew')
        self.ent_cost.insert(0, "0")
        self.ent_cost.bind('<Return>', lambda e: btn_reg.focus_set())
        
        # Row 8
        btn_reg = ttk.Button(reg_frame, text="자재 등록", command=self.register_material)
        btn_reg.grid(row=8, column=0, columnspan=4, pady=10)
        

        
        # Side 2: Transaction Entry
        trans_reg_container = ttk.Frame(self.inout_top_paned)
        self.inout_top_paned.add(trans_reg_container, weight=1)
        
        # Frame for In/Out Transaction (Directly packed)
        trans_frame = ttk.LabelFrame(trans_reg_container, text="입출고 기록")
        trans_frame.pack(fill='both', expand=True, padx=10, pady=5)
        trans_frame.grid_columnconfigure(1, weight=1)
        trans_frame.grid_columnconfigure(3, weight=1)

        # Bottom frame for history (Occupies full width at bottom)
        history_container = ttk.Frame(self.inout_paned)
        self.inout_paned.add(history_container, weight=2)
        
        ttk.Label(trans_frame, text="자재 선택:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        mat_sel_container = ttk.Frame(trans_frame)
        mat_sel_container.grid(row=0, column=1, padx=5, pady=2, columnspan=3, sticky='ew')
        
        self.cb_material = ttk.Combobox(mat_sel_container, width=45, font=('Malgun Gothic', 10))
        self.cb_material.pack(side='left', fill='x', expand=True)
        self.cb_material.bind('<<ComboboxSelected>>', self.on_material_selected)
        self.cb_material.bind('<Return>', lambda e: self.cb_type.focus_set())
        
        # [MODERN] Place the search button INSIDE the combobox
        btn_mat_search = tk.Button(self.cb_material, text="🔍", font=('Arial', 8), 
                                    bd=0, bg='white', cursor='hand2',
                                    command=self.open_material_search_dialog)
        # Offset x=-20 to avoid overlapping the combobox arrow
        btn_mat_search.place(relx=1.0, x=-20, rely=0.5, anchor='e', width=18, height=18)
        
        # Register autocomplete for material selection
        self.enable_autocomplete(self.cb_material, values_list=[]) # Initial empty, will be updated
        self.update_material_combo()
        
        # Model List display - Moved to row 4 for narrower fit
        ttk.Label(trans_frame, text="관련 모델명:").grid(row=4, column=0, padx=5, pady=2, sticky='nw')
        self.list_models = tk.Listbox(trans_frame, height=3, width=42, font=('Malgun Gothic', 9))
        self.list_models.grid(row=4, column=1, columnspan=3, padx=5, pady=2, sticky='ew')
        
        # Add scrollbar for model list
        model_vsb = ttk.Scrollbar(trans_frame, orient="vertical", command=self.list_models.yview)
        model_vsb.grid(row=4, column=4, sticky='ns', pady=2)
        self.list_models.config(yscrollcommand=model_vsb.set)
        
        ttk.Label(trans_frame, text="구분:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.cb_type = ttk.Combobox(trans_frame, values=["IN", "OUT"], state="readonly", width=15)
        self.cb_type.grid(row=1, column=1, padx=5, pady=2, sticky='ew')
        self.cb_type.set("OUT")
        self.cb_type.bind('<Return>', lambda e: self.ent_qty.focus_set())
        
        ttk.Label(trans_frame, text="수량:").grid(row=1, column=2, padx=5, pady=2, sticky='w')
        self.ent_qty = ttk.Entry(trans_frame, width=30)
        self.ent_qty.grid(row=1, column=3, padx=5, pady=2, sticky='ew')
        self.ent_qty.bind('<Return>', lambda e: self.cb_trans_site.focus_set())
        
        ttk.Label(trans_frame, text="현장:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.cb_trans_site = ttk.Combobox(trans_frame, width=28, values=self.sites)
        self.cb_trans_site.grid(row=2, column=1, padx=5, pady=2, sticky='ew')
        self.cb_trans_site.bind('<FocusOut>', lambda e: self.auto_save_to_list(e, self.cb_trans_site, self.sites, 'sites'))
        self.cb_trans_site.bind('<Return>', lambda e: self._on_trans_site_return(e))
        
        ttk.Label(trans_frame, text="창고:").grid(row=2, column=2, padx=5, pady=2, sticky='w')
        self.cb_warehouse = ttk.Combobox(trans_frame, width=28, values=self.warehouses)
        self.cb_warehouse.grid(row=2, column=3, padx=5, pady=2, sticky='ew')
        self.cb_warehouse.bind('<FocusOut>', lambda e: self.auto_save_to_list(e, self.cb_warehouse, self.warehouses, 'warehouses'))
        self.cb_warehouse.bind('<Return>', lambda e: self._on_warehouse_return(e))
        
        ttk.Label(trans_frame, text="담당자:").grid(row=3, column=0, padx=5, pady=2, sticky='w')
        self.ent_user = ttk.Combobox(trans_frame, width=28, values=getattr(self, 'users', []))
        self.ent_user.grid(row=3, column=1, padx=5, pady=2, sticky='ew')
        self.ent_user.bind('<FocusOut>', lambda e: self.auto_save_to_list(e, self.ent_user, self.users, 'users'))
        self.ent_user.bind('<Return>', lambda e: self._on_user_return(e))
        
        ttk.Label(trans_frame, text="비고:").grid(row=3, column=2, padx=5, pady=2, sticky='w')
        self.ent_note = ttk.Entry(trans_frame, width=30)
        self.ent_note.grid(row=3, column=3, padx=5, pady=2, sticky='ew')
        self.ent_note.bind('<Return>', lambda e: btn_trans.focus_set())
        
        btn_trans = ttk.Button(trans_frame, text="기록 저장", command=self.add_transaction)
        btn_trans.grid(row=5, column=0, columnspan=4, pady=10)
        
        # Frame for displaying transaction history
        history_frame = ttk.LabelFrame(history_container, text="최근 입출고 내역")
        history_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Filter/Control frame for history
        history_ctrl_frame = ttk.Frame(history_frame)
        history_ctrl_frame.pack(fill='x', padx=5, pady=2)
        
        # First row of filters
        filter_row1 = ttk.Frame(history_ctrl_frame)
        filter_row1.pack(fill='x', pady=2)
        
        btn_del_trans = ttk.Button(filter_row1, text="선택 항목 삭제", command=self.delete_transaction_entry)
        btn_del_trans.pack(side='left', padx=5)
        
        btn_export_trans = ttk.Button(filter_row1, text="엑셀 내보내기", command=self.export_transaction_history)
        btn_export_trans.pack(side='left', padx=5)

        btn_refresh_trans = ttk.Button(filter_row1, text="입출고 내역 새로고침", command=self.refresh_inout_history)
        btn_refresh_trans.pack(side='left', padx=5)

        # [NEW] Cleanup button for automatic transactions (Equipment only)
        btn_cleanup_trans = ttk.Button(filter_row1, text="🧹 장비류 자동기록 클린업", command=self.cleanup_auto_transactions)
        btn_cleanup_trans.pack(side='left', padx=10)
        
        # Second row of filters
        filter_row2 = ttk.Frame(history_ctrl_frame)
        filter_row2.pack(fill='x', pady=2)
        
        ttk.Label(filter_row2, text="품목명 필터:").pack(side='left', padx=5)
        self.cb_trans_filter_mat = ttk.Combobox(filter_row2, width=35, state="readonly")
        self.cb_trans_filter_mat.pack(side='left', padx=5)
        self.cb_trans_filter_mat.bind('<<ComboboxSelected>>', lambda e: self.update_transaction_view())
        
        ttk.Label(filter_row2, text="현장 필터:").pack(side='left', padx=(20, 5))
        self.cb_trans_filter_site = ttk.Combobox(filter_row2, width=12, state="readonly")
        self.cb_trans_filter_site.pack(side='left', padx=5)
        self.cb_trans_filter_site.bind('<<ComboboxSelected>>', lambda e: self.update_transaction_view())
        
        # Third row of filters
        filter_row3 = ttk.Frame(history_ctrl_frame)
        filter_row3.pack(fill='x', pady=2)
        
        ttk.Label(filter_row3, text="차량번호 필터:").pack(side='left', padx=5)
        self.cb_trans_filter_vehicle = ttk.Combobox(filter_row3, width=12, state="readonly")
        self.cb_trans_filter_vehicle.pack(side='left', padx=5)
        self.cb_trans_filter_vehicle.bind('<<ComboboxSelected>>', lambda e: self.update_transaction_view())
        self.cb_trans_filter_vehicle.bind('<F5>', lambda e: (self.refresh_inout_history(), 'break')[1])
        
        # Treeview for history
        tree_scroll_frame = ttk.Frame(history_frame)
        tree_scroll_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        inout_vsb = ttk.Scrollbar(tree_scroll_frame, orient="vertical")
        inout_hsb = ttk.Scrollbar(tree_scroll_frame, orient="horizontal")
        
        # [RESTRUCTURED] Optimized for NDT consumables: Focus on RT, PT, MT quantities
        columns = ('날짜', '현장', '구분', '품목명', '수량', '세척제', '침투제', '현상제', '백색페인트', '흑색자분', 'SN', '규격', '창고', '담당자', '비고', '검사비')
        self.inout_tree = ttk.Treeview(tree_scroll_frame, columns=columns, show='headings', height=10,
                                       yscrollcommand=inout_vsb.set, xscrollcommand=inout_hsb.set)
        
        inout_vsb.config(command=self.inout_tree.yview)
        inout_hsb.config(command=self.inout_tree.xview)
        
        # [NEW] Default Column Widths for NDT-focused View
        col_widths = {
            '날짜': 130, '현장': 120, '구분': 60, '품목명': 160, 
            '수량': 70, '세척제': 60, '침투제': 60, '현상제': 60, '백색페인트': 80, '흑색자분': 80,
            'SN': 100, '규격': 100, '창고': 80, '담당자': 80, '비고': 150, '검사비': 90
        }
        for col in columns:
            self.inout_tree.heading(col, text=col)
            width = col_widths.get(col, 100)
            self.inout_tree.column(col, width=width, minwidth=50, stretch=False, anchor='center')
        
        # [FIX] Select only the core NDT-relevant columns for initial display
        visible_cols = ('날짜', '현장', '구분', '품목명', '수량', '세척제', '침투제', '현상제', '백색페인트', '흑색자분', '비고')
        self.inout_tree['displaycolumns'] = visible_cols
        self.enable_tree_column_drag(self.inout_tree, context_menu_handler=lambda e: self._show_generic_tree_heading_context_menu(e, self.inout_tree))
        self.inout_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config(), add='+')
        self.inout_tree.bind('<F5>', lambda e: (self.refresh_inout_history(), 'break')[1])

        self.inout_tree.grid(row=0, column=0, sticky='nsew')
        inout_vsb.grid(row=0, column=1, sticky='ns')
        inout_hsb.grid(row=1, column=0, sticky='ew')
        
        tree_scroll_frame.grid_rowconfigure(0, weight=1)
        tree_scroll_frame.grid_columnconfigure(0, weight=1)
        
        # Initial populate
        self.update_transaction_view()
        
        # Set initial sash position for better balance
        self.inout_paned.after(200, self._ensure_inout_sash_visibility)

    def _ensure_inout_sash_visibility(self):
        """Ensure the inout sash position is properly set"""
        try:
            if hasattr(self, 'inout_paned'):
                self.inout_paned.update_idletasks()
                total_h = self.inout_paned.winfo_height()
                if total_h > 200:
                    # Set vertical sash to 40% of total height (registration area smaller)
                    new_pos = int(total_h * 0.4)
                    self.inout_paned.sashpos(0, new_pos)
                    
                    # Set horizontal sash for side-by-side
                    total_w = self.inout_top_paned.winfo_width()
                    if total_w > 200:
                        self.inout_top_paned.sashpos(0, int(total_w * 0.45))
                    print(f"Set inout sash to {new_pos} (total height: {total_h})")
        except Exception as e:
            print(f"Error ensuring inout sash visibility: {e}")


    def get_material_display_name(self, mat_id):
        """Get formatted material name as '품목명 (SN: SN번호) - 규격'"""
        # [FIX] Handle NaN IDs gracefully to prevent "NAN" display
        if pd.isna(mat_id) or str(mat_id).lower().strip() == 'nan':
            return "미지정 품목"
            
        if self.materials_df.empty:
            return f"ID: {mat_id}"
            
        mat_row = self.materials_df[self.materials_df['MaterialID'] == mat_id]
        if mat_row.empty:
            # [NEW] Handle non-numeric IDs (raw names for PAUT) and orphans gracefully
            return str(mat_id)
            
        mat = mat_row.iloc[0]
        name = mat['품목명']
        sn = mat.get('SN', '')
        spec = mat.get('규격', '')
        
        # Build display string: 품목명-SN (모델명) [규격]
        display = name
        if sn and pd.notna(sn) and str(sn).strip():
            display = f"{name}-{str(sn).strip()}"
            
        model_name = mat.get('모델명', '')
        model_name_val = str(model_name).strip()
        if model_name_val and pd.notna(model_name) and model_name_val != str(name).strip():
            display = f"{display} ({model_name_val})"
        
        if spec and pd.notna(spec) and str(spec).strip():
            display = f"{display} [{str(spec).strip()}]"
            
        return display

    def get_material_name_only(self, mat_id):
        """Get only the 품목명 for a material"""
        if self.materials_df.empty:
            return f"(미등록) {mat_id}"
        
        target_id = self.normalize_id(mat_id)
        # Match using normalized IDs to handle formatting variations
        mask = self.materials_df['MaterialID'].apply(self.normalize_id) == target_id
        mat_row = self.materials_df[mask]
        
        if mat_row.empty:
            return f"(미등록) {mat_id}"
        return str(mat_row.iloc[0].get('품목명', f"(미등록) {mat_id}"))

    def get_material_info(self, mat_id):
        """Get all material columns as a dictionary"""
        if self.materials_df.empty:
            return {}
        
        target_id = self.normalize_id(mat_id)
        mask = self.materials_df['MaterialID'].apply(self.normalize_id) == target_id
        mat_row = self.materials_df[mask]
        
        if mat_row.empty:
            return {}
        return mat_row.iloc[0].to_dict()
        
        # Add specification if exists
        if spec and pd.notna(spec) and str(spec).strip():
            display += f" - {spec}"
            
        return display

    def update_material_combo(self):
        """Update both In/Out and Daily Usage material comboboxes with unified list"""
        mat_list = []
        if not self.materials_df.empty:
            # Create list with unified display format
            # [STRICT] Robust check for Active status
            active_mask = pd.to_numeric(self.materials_df.get('Active', 1), errors='coerce').fillna(1) > 0.5
            active_df = self.materials_df[active_mask]
            
            for _, mat in active_df.iterrows():
                display = self.get_material_display_name(mat['MaterialID'])
                mat_list.append(display)
        
        # [NEW] Pre-calculate inactive items for efficient filtering (STRICT)
        inactive_ids = set()
        inactive_names = set()
        if not self.materials_df.empty:
            # Use strict numeric check
            active_vals = pd.to_numeric(self.materials_df.get('Active', 1), errors='coerce').fillna(1)
            for idx, m in self.materials_df.iterrows():
                if active_vals.loc[idx] < 0.5:
                    inactive_ids.add(self.normalize_id(m.get('MaterialID')))
                    inactive_names.add(str(m.get('품목명', '')).strip())

        # [NEW] Also collect historical material names from daily usage history (Only if still Active)
        history_materials = []
        if not self.daily_usage_df.empty and 'MaterialID' in self.daily_usage_df.columns:
            unique_mat_ids = self.daily_usage_df['MaterialID'].dropna().unique()
            for mat_id in unique_mat_ids:
                str_id = self.normalize_id(mat_id)
                if str_id in inactive_ids:
                    continue
                name = self.get_material_display_name(mat_id)
                if name: history_materials.append(name)
        
        # Merge database items with centralized films list and historical materials, unique and sort
        all_vals = list(set([str(m) for m in mat_list + self.carestream_films + history_materials if pd.notna(m) and str(m).strip()]))
        all_vals.sort()
        
        # Update ComboBoxes
        self.cached_material_list = all_vals
        if hasattr(self, 'cb_material'):
            self.cb_material['values'] = all_vals
        if hasattr(self, 'cb_daily_material'):
            if isinstance(self.cb_daily_material, ttk.Combobox):
                self.cb_daily_material['values'] = all_vals
                try:
                    max_chars = max((len(str(v)) for v in all_vals), default=12)
                    self.cb_daily_material.configure(width=min(40, max(22, max_chars + 2)))
                except Exception:
                    pass
            
        if hasattr(self, 'cb_daily_equip'):
            # [NEW] Collect history of equipment names from daily usage records (Only if Active)
            history_equip = []
            if not self.daily_usage_df.empty and '장비명' in self.daily_usage_df.columns:
                raw_history = self.daily_usage_df['장비명'].dropna().unique().tolist()
                for eq in raw_history:
                    eq_str = str(eq).strip()
                    if not eq_str or eq_str in inactive_names:
                        continue
                    history_equip.append(eq_str)
                
            # [NEW] Filter: Only include actual equipment (non-consumables) in the equipment dropdown
            equip_only_vals = []
            for v in all_vals:
                base_name = v.split('[')[0].split('-')[0].strip()
                if not self._is_consumable_material(base_name, ''):
                    equip_only_vals.append(v)
                
            # Combine custom equipments, history, and the filtered equipment list
            combined_equip = list(set([str(e).strip() for e in self.equipments + history_equip + equip_only_vals if pd.notna(e) and str(e).strip()]))
            combined_equip.sort()
            
            # [FIX] Only set values if it's a Combobox
            if isinstance(self.cb_daily_equip, ttk.Combobox):
                self.cb_daily_equip['values'] = combined_equip
            
            try:
                max_chars = max((len(str(v)) for v in combined_equip), default=12)
                self.cb_daily_equip.configure(width=min(40, max(22, max_chars + 2)))
            except Exception:
                pass
            self.equipment_suggestions = combined_equip
        
        self.materials_display_list = all_vals
        
        # [NEW] Calculate consumable-only list for Inventory Status filters (Exclude PAUT, etc.)
        self.consumable_display_list = []
        for val in all_vals:
            # Extract the base name (before spec '-' or ID '[')
            base_name = val.split('[')[0].split('-')[0].strip()
            if self._is_consumable_material(base_name, ''):
                self.consumable_display_list.append(val)
        self.consumable_display_list.sort()


        if hasattr(self, 'cb_trans_filter_mat'):
            self.cb_trans_filter_mat['values'] = ["전체"] + all_vals
            if not self.cb_trans_filter_mat.get():
                self.cb_trans_filter_mat.set("전체")
        
        if hasattr(self, 'cb_trans_filter_site'):
            self.cb_trans_filter_site['values'] = ["전체"] + sorted(self.sites)
            if not self.cb_trans_filter_site.get():
                self.cb_trans_filter_site.set("전체")
        
        if hasattr(self, 'cb_trans_filter_vehicle'):
            # Get unique vehicle numbers from transactions_df
            if '차량번호' in self.transactions_df.columns:
                vehicle_numbers = self.transactions_df['차량번호'].dropna().astype(str).str.strip().unique()
                vehicle_list = sorted([v for v in vehicle_numbers if v and v != 'nan'])
            else:
                vehicle_list = []
            
            # Add sample vehicle numbers for testing if no data exists
            if not vehicle_list:
                vehicle_list = ['12가1234', '34나5678', '89다9012']
            
            self.cb_trans_filter_vehicle['values'] = ["전체"] + vehicle_list
            if not self.cb_trans_filter_vehicle.get():
                self.cb_trans_filter_vehicle.set("전체")

    def _get_material_candidates(self, include_all=False):
        vals = []
        if hasattr(self, 'materials_display_list'):
            vals.extend(getattr(self, 'materials_display_list', []) or [])
        if hasattr(self, 'cb_daily_material'):
            pass # Removed reading from widget to prevent recursive filtering issues

        cleaned = sorted(set(str(v).strip() for v in vals if str(v).strip() and str(v).strip() != '전체' and not str(v).strip().startswith('ID: ')))
        return (['전체'] + cleaned) if include_all else cleaned

    def _get_history_material_candidates(self, include_all=False):
        """Get material names from both current inventory and history for selection"""
        raw_materials = set()
        
        # 1. Add all active materials from stock master (Primary Source)
        if not self.materials_df.empty:
            for _, mat in self.materials_df.iterrows():
                if mat.get('Active', 1) == 1:
                    name = self.get_material_display_name(mat['MaterialID'])
                    if name: raw_materials.add(name)
        
        # 2. Add materials exclusively from daily usage history (Secondary Source)
        if not self.daily_usage_df.empty and 'MaterialID' in self.daily_usage_df.columns:
            unique_mat_ids = self.daily_usage_df['MaterialID'].dropna().unique()
            for mat_id in unique_mat_ids:
                name = self.get_material_display_name(mat_id)
                if name: raw_materials.add(name)
        
        cleaned = sorted(list(raw_materials))
        return (['전체'] + cleaned) if include_all else cleaned

    def _get_equipment_candidates(self, include_all=False):
        """Get unified equipment suggestions from master list and history"""
        raw_equip = set()
        
        # 1. Base equipment list
        if hasattr(self, 'equipments'):
            for e in getattr(self, 'equipments', []):
                if e and str(e).strip(): raw_equip.add(str(e).strip())
                
        # 2. Dynamic suggestions
        if hasattr(self, 'equipment_suggestions'):
            for e in getattr(self, 'equipment_suggestions', []):
                if e and str(e).strip(): raw_equip.add(str(e).strip())
                
        # 3. History
        if hasattr(self, 'daily_usage_df') and not self.daily_usage_df.empty and '장비명' in self.daily_usage_df.columns:
            unique_equip = self.daily_usage_df['장비명'].dropna().astype(str).str.strip().unique()
            for e in unique_equip:
                if e: raw_equip.add(e)

        cleaned = sorted([e for e in raw_equip if e != '전체'])
        return (['전체'] + cleaned) if include_all else cleaned

    def _get_inspection_item_candidates(self):
        """Get candidates for Inspection Item (Inspection Item Name)"""
        items = set(getattr(self, 'test_items', []))
        if not self.daily_usage_df.empty and '검사품명' in self.daily_usage_df.columns:
            hist = self.daily_usage_df['검사품명'].dropna().astype(str).str.strip().unique()
            for h in hist: 
                if h: items.add(h)
        return sorted(list(items))

    def _get_applied_code_candidates(self):
        """Get candidates for Applied Code"""
        codes = set(getattr(self, 'applied_codes', []))
        if not self.daily_usage_df.empty and '적용코드' in self.daily_usage_df.columns:
            hist = self.daily_usage_df['적용코드'].dropna().astype(str).str.strip().unique()
            for h in hist: 
                if h: codes.add(h)
        return sorted(list(codes))

    def _apply_combobox_word_suggest(self, combobox, source_values, open_dropdown=False):
        if not combobox or not isinstance(combobox, ttk.Combobox):
            return

        text_raw = combobox.get()
        text = text_raw.strip()
        had_focus = False
        try:
            # More robust focus check
            curr_focus = self.root.focus_get()
            had_focus = (curr_focus == combobox or str(curr_focus).startswith(str(combobox)))
        except Exception:
            had_focus = False
        
        try:
            cursor_pos = combobox.index(tk.INSERT)
        except Exception:
            cursor_pos = len(text_raw)
            
        source = [str(v).strip() for v in (source_values or []) if str(v).strip()]

        if not text:
            filtered = source
        else:
            q = text.lower()
            filtered = [v for v in source if q in v.lower()]

        # [FIX] Only update if values changed to avoid recursion/flicker
        values_changed = list(combobox['values']) != list(filtered)
        if values_changed:
            combobox['values'] = filtered

        # Always ensure state is reconciled if it had focus
        if had_focus:
            def _restore_state():
                try:
                    if self.root.focus_get() != combobox:
                        combobox.focus_set()
                    combobox.set(text_raw)
                    combobox.icursor(cursor_pos)
                    combobox.selection_clear()
                except Exception:
                    pass

            _restore_state()
            # Double tap with after_idle to ensure it sticks after the internal Tk events
            try: combobox.after_idle(_restore_state)
            except Exception: pass
        
        # [REVISION] Use 'Alt-Down' for more stable dropdown posting
        if open_dropdown and had_focus and filtered and len(text) > 0:
            try:
                # [STABILITY] Don't re-open if just selected
                if getattr(combobox, '_just_selected', False):
                    return
                # Alt-Down is the standard shortcut to post the dropdown in many OS/Toolkits
                combobox.event_generate('<Alt-Down>')
                # Re-verify cursor immediately after posting
                combobox.icursor(cursor_pos)
            except:
                pass

    def _bind_combobox_word_suggest(self, combobox, source_getter):
        if not combobox:
            return

        def _on_interaction(e=None, open=True):
            # Skip for navigation keys which have native behaviors
            if e is not None and hasattr(e, 'keysym'):
                if e.keysym in {
                    'Left', 'Right', 'Tab', 'Up', 'Down', 'Return', 'Escape',
                    'Shift_L', 'Shift_R', 'Control_L', 'Control_R',
                    'Alt_L', 'Alt_R', 'Caps_Lock'
                }:
                    return
                # [NEW] Skip 'Process' keysym used by Korean IME during composition
                if e.keysym == 'Process':
                    return
            
            # [FIX] Debounce: Wait 300ms before showing suggestions to avoid interrupting continuous typing (especially for IME)
            if hasattr(combobox, '_suggest_after_id'):
                combobox.after_cancel(combobox._suggest_after_id)
            
            def _do_suggest():
                if combobox.winfo_exists():
                    self._apply_combobox_word_suggest(combobox, source_getter(), open_dropdown=open)
            
            combobox._suggest_after_id = self.root.after(300, _do_suggest)

        combobox.bind('<KeyRelease>', lambda e: _on_interaction(e, open=True), add='+')
        combobox.bind('<FocusIn>', lambda e: _on_interaction(e, open=False), add='+')
        
        # [NEW] Guard against "sticky" dropdown after selection
        def _on_selected(e=None):
            combobox._just_selected = True
            def _reset(): combobox._just_selected = False
            combobox.after(500, _reset)
            
        combobox.bind('<<ComboboxSelected>>', _on_selected, add='+')

    def update_registration_combos(self):
        """Update registration comboboxes with unique values from database and centralized list"""
        # 1. Update registration fields from database
        fields = {}
        for key, attr in [
            ('회사코드', 'cb_co_code'),
            ('관리품번', 'cb_eq_code'),
            ('품목명', 'cb_item_name'),
            ('품목군코드', 'cb_class'),
            ('규격', 'cb_spec'),
            ('관리단위', 'cb_unit'),
            ('공급업체', 'cb_supplier'),
            ('제조사', 'cb_mfr'),
            ('제조국', 'cb_origin')
        ]:
            if hasattr(self, attr):
                fields[key] = getattr(self, attr)
        
        # Store lists for autocomplete (Main Catalog + History)
        self.co_code_list = []
        self.eq_code_list = []
        self.item_name_list = []
        self.class_list = []
        self.spec_list = []
        self.unit_list = []
        self.supplier_list = []
        self.mfr_list = []
        self.origin_list = []
        self.sn_list = []
        self.model_list = []

        attr_mapping = {
            '회사코드': 'co_code_list',
            '관리품번': 'eq_code_list',
            '품목명': 'item_name_list',
            '품목군코드': 'class_list',
            '규격': 'spec_list',
            '관리단위': 'unit_list',
            '공급업체': 'supplier_list',
            '제조사': 'mfr_list',
            '제조국': 'origin_list',
            'SN': 'sn_list',
            '모델명': 'model_list'
        }

        for col, list_attr in attr_mapping.items():
            vals = []
            if not self.materials_df.empty and col in self.materials_df.columns:
                # [STRICT] Robust check for Active status (numeric comparison)
                # Ensure Active is treated as numeric, defaulting to 1 (Active)
                active_mask = pd.to_numeric(self.materials_df.get('Active', 1), errors='coerce').fillna(1) > 0.5
                active_df = self.materials_df[active_mask]
                unique_vals = active_df[col].dropna().unique()
                vals = sorted([str(v).strip() for v in unique_vals if v and str(v).strip()])
            
            # Update instance attributes for autocomplete
            if col == '관리단위':
                # Merge DB units with the managed daily_units
                self.unit_list = list(dict.fromkeys(self.daily_units + vals))
                if hasattr(self, 'cb_daily_unit'):
                    self.cb_daily_unit['values'] = self.unit_list
            else:
                setattr(self, list_attr, vals)
            
            # 1. Update registration fields (without "전체")
            if col in fields:
                fields[col]['values'] = vals

            # 2. Update Stock View filters (with "전체")
            # [NEW] For Stock View filters, also exclude non-consumable equipment to match the list view
            filter_map = {
                '회사코드': getattr(self, 'cb_filter_co', None),
                '품목군코드': getattr(self, 'cb_filter_class', None),
                '제조사': getattr(self, 'cb_filter_mfr', None),
                '품목명': getattr(self, 'cb_filter_name', None),
                'SN': getattr(self, 'cb_filter_sn', None),
                '모델명': getattr(self, 'cb_filter_model', None),
                '관리품번': getattr(self, 'cb_filter_eq', None)
            }
            if col in filter_map and filter_map[col] is not None:
                combo = filter_map[col]
                
                # Filter vals for consumables if it's the item name or model
                c_vals = vals
                if col in ['품목명', '모델명', '관리품번']:
                    c_vals = []
                    for v in vals:
                        # Heuristic: Check if any item with this name/model is a consumable
                        if col == '품목명':
                            if self._is_consumable_material(v, ''): c_vals.append(v)
                        elif col == '모델명':
                            # Check if at least one material with this model is a consumable
                            m_rows = self.materials_df[self.materials_df['모델명'].astype(str).str.strip() == v]
                            if not m_rows.empty and self._is_consumable_material(m_rows.iloc[0].get('품목명', ''), ''):
                                c_vals.append(v)
                        else:
                            c_vals.append(v)
                
                combo['values'] = ["전체"] + c_vals
                if not combo.get():
                    combo.set("전체")
        
        # Add Carestream film options to 품목명 (Registration tab only)
        if hasattr(self, 'cb_item_name'):
            all_mat_vals = sorted(list(set([str(mat) for mat in self.item_name_list + self.carestream_films if pd.notna(mat) and str(mat).strip()])))
            self.item_name_list[:] = all_mat_vals
            self.cb_item_name['values'] = self.item_name_list
            
        # [FIX] Do NOT overwrite cb_filter_name values here; it was already set with consumable filtering above.
        # Just ensure the autocomplete list is refreshed.
        if hasattr(self, 'apply_autocomplete_to_all_comboboxes'):
            self.apply_autocomplete_to_all_comboboxes()
        
        # 3. Final view update now that filters are set to "전체"
        import threading
        self.root.after(100, self.update_stock_view)

    def register_material(self):
        co_code = self.cb_co_code.get()
        item_name = self.cb_item_name.get()
        eq_code = self.cb_eq_code.get()
        sn = self.ent_sn.get()
        classification = self.cb_class.get()
        spec = self.cb_spec.get()
        unit = self.cb_unit.get()
        supplier = self.cb_supplier.get()
        manufacturer = self.cb_mfr.get()
        origin = self.cb_origin.get()
        
        try:
            init_stock = float(self.ent_init.get())
            reorder_point = float(self.ent_reorder.get())
            price_val = float(self.ent_price.get() if self.ent_price.get() else 0)
            cost_val = float(self.ent_cost.get() if self.ent_cost.get() else 0)
        except ValueError:
            messagebox.showwarning("입력 오류", "재고, 단가, 원가 정보는 숫자여야 합니다.")
            return
            
        if not item_name:
            messagebox.showwarning("입력 오류", "품목명을 입력해주세요.")
            return
        
        # Generate MaterialID
        if self.materials_df.empty:
            mat_id = 1
        else:
            mat_id = self.materials_df['MaterialID'].max() + 1
        
        # Extract SN from Model Name (Now using newly added model name combobox)
        model_name = self.cb_model.get().strip() if hasattr(self, 'cb_model') else ''
        new_model, new_sn = self.extract_sn_from_model(model_name, sn)
        # Note: In register_material, model_name is currently not an input field, but SN is.
        # If the user adds model name to registration later, this will handle it.
        
        new_row = {
            'MaterialID': mat_id,
            '회사코드': co_code,
            '관리품번': eq_code,
            '품목명': item_name,
            'SN': new_sn,
            '창고': '',
            '모델명': new_model,
            '규격': spec,
            '품목군코드': classification,
            '공급업체': supplier,
            '제조사': manufacturer,
            '제조국': origin,
            '가격': price_val,
            '원가': cost_val,
            '관리단위': unit if unit else 'EA',
            '수량': init_stock,
            '재고하한': reorder_point
        }
        
        self.materials_df = pd.concat([self.materials_df, pd.DataFrame([new_row])], ignore_index=True)
        self.save_data()
        self.update_material_combo()
        self.update_registration_combos()
        self.update_stock_view()
        messagebox.showinfo("완료", f"'{item_name}' 자재가 등록되었습니다.")
        
        # Clear entries
        self.cb_co_code.set('')
        self.cb_eq_code.set('')
        self.cb_item_name.set('')
        self.ent_sn.delete(0, tk.END)
        self.cb_class.set('')
        self.cb_spec.set('')
        self.cb_unit.set('')
        self.cb_supplier.set('')
        self.cb_mfr.set('')
        self.cb_origin.set('')
        self.ent_reorder.delete(0, tk.END)
        self.ent_reorder.insert(0, "0")
        self.ent_init.delete(0, tk.END)
        self.ent_init.insert(0, "0")
        # Clear additional fields
        if hasattr(self, 'cb_model'):
            self.cb_model.set('')
        if hasattr(self, 'ent_price'):
            self.ent_price.delete(0, tk.END)
        if hasattr(self, 'ent_cost'):
            self.ent_cost.delete(0, tk.END)

    def add_transaction(self):
        """Record an IN or OUT transaction"""
        try:
            mat_selection = self.cb_material.get().strip()
            t_type = self.cb_type.get()
            user = self.ent_user.get().strip()
            
            if not mat_selection:
                messagebox.showwarning("입력 오류", "자재를 선택해주세요.")
                return
            if not t_type:
                messagebox.showwarning("입력 오류", "구분(입고/출고)을 선택해주세요.")
                return
                
            try:
                qty_str = self.ent_qty.get().strip()
                if not qty_str:
                    messagebox.showwarning("입력 오류", "수량을 입력해주세요.")
                    return
                qty = float(qty_str)
            except ValueError:
                messagebox.showwarning("입력 오류", "수량은 숫자여야 합니다.")
                return
                
            note = self.ent_note.get().strip()
            
            # Extract pure material name from selection
            mat_name = mat_selection
            if " - " in mat_name: mat_name = mat_name.split(" - ")[0]
            if " (SN: " in mat_name: mat_name = mat_name.split(" (SN: ")[0]
            pure_mat_name = mat_name.strip()
            
            # Find MaterialID
            # Ensure MaterialID is treated consistently
            mat_rows = self.materials_df[self.materials_df['품목명'] == pure_mat_name]
            
            if mat_rows.empty:
                # If exact match fails, try a case-insensitive or stripped match
                mat_rows = self.materials_df[self.materials_df['품목명'].str.strip() == pure_mat_name]
            
            if mat_rows.empty:
                # [NEW] Ask to register new material instead of just showing error
                confirm = messagebox.askyesno("신규 자재", f"'{pure_mat_name}'은(는) 등록되지 않은 자재입니다. 신규 자재로 등록하고 진행할까요?")
                if confirm:
                    mat_id = self.register_new_material(pure_mat_name)
                else:
                    return
            else:
                mat_id = mat_rows['MaterialID'].values[0]
            
            # Update Warehouse in materials_df
            warehouse = str(self.cb_warehouse.get()).strip()
            if warehouse:
                # Type-safe assignment
                if '창고' in self.materials_df.columns:
                    mask = self.materials_df['MaterialID'] == mat_id
                    if mask.any():
                        self.materials_df.loc[mask, '창고'] = warehouse
            
            # Create transaction record
            new_trans = {
                'Date': datetime.datetime.now(),
                'MaterialID': mat_id,
                'Type': t_type,
                'Quantity': qty if t_type == 'IN' else -qty,
                'Note': note,
                'User': user,
                'Site': self.cb_trans_site.get() if hasattr(self, 'cb_trans_site') else '',
                '차량번호': '',
                '주행거리': '',
                '차량점검': '',
                '차량비고': ''
            }
            
            # Add to dataframe
            self.transactions_df = pd.concat([self.transactions_df, pd.DataFrame([new_trans])], ignore_index=True)
            
            # Save all data (force sync)
            self.save_data()
            
            # Check if it was really added (for debug feedback)
            last_count = len(self.transactions_df)
            
            # Refresh views
            self.update_stock_view()
            self.update_transaction_view()
            self.update_material_combo() 
            
            # Auto-save site/user to lists
            site_value = new_trans['Site']
            if site_value and site_value not in self.sites:
                self.sites.append(site_value)
                self.sites.sort()
                self.save_tab_config()
            
            if user and user not in self.users:
                self.users.append(user)
                self.users.sort()
                self.save_tab_config()
            
            # Success feedback
            messagebox.showinfo("완료", f"{pure_mat_name} {t_type} 처리되었습니다.\n(전체 기록 수: {last_count}개)")
            
            # Clear UI fields
            self.ent_qty.delete(0, tk.END)
            self.ent_note.delete(0, tk.END)
            if hasattr(self, 'ent_user') and hasattr(self.ent_user, 'set'):
                self.ent_user.set('')
            elif hasattr(self, 'ent_user'):
                self.ent_user.delete(0, tk.END)
            
            if hasattr(self, 'cb_warehouse'): self.cb_warehouse.set('')
            if hasattr(self, 'cb_trans_site'): self.cb_trans_site.set('')
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            self.show_error_dialog("저장 오류", f"기록 저장 중 기술적인 오류가 발생했습니다:\n{e}\n\n상세 정보:\n{error_details}")
        
        # Ensure view is updated regardless
        self.update_transaction_view()

    def update_transaction_view(self):
        """Unified transaction view that combines manual stock transactions and site usage."""
        try:
            # 0. Helper Functions (Global normalize_id is now used)
            def _f(val):
                if val is None or pd.isna(val): return 0.0
                try: 
                    s = str(val).replace(',', '').strip()
                    return float(s) if s else 0.0
                except: return 0.0
                
            def _q_str(val):
                try:
                    v = abs(float(str(val).replace(',', '').strip()))
                    return f"{v:g}" if v > 0 else ""
                except: return ""

            # 1. Clear existing view
            for item in self.inout_tree.get_children():
                self.inout_tree.delete(item)
            
            # 2. Prepare Transaction Records
            df_trans = pd.DataFrame()
            active_ids = []
            if not self.materials_df.empty and 'MaterialID' in self.materials_df.columns:
                # Get list of MaterialIDs that are NOT deleted (Active != 0)
                active_mask = (self.materials_df.get('Active', 1) != 0)
                active_ids = [self.normalize_id(x) for x in self.materials_df[active_mask]['MaterialID'].dropna().tolist()]
                
            if not self.transactions_df.empty:
                df_trans = self.transactions_df.copy()
                # Column Compatibility
                if 'Site' not in df_trans.columns and '현장' in df_trans.columns: df_trans['Site'] = df_trans['현장']
                if 'Date' not in df_trans.columns and '날짜' in df_trans.columns: df_trans['Date'] = df_trans['날짜']
                if 'Date' not in df_trans.columns and '일자' in df_trans.columns: df_trans['Date'] = df_trans['일자']
                
                df_trans['Date'] = pd.to_datetime(df_trans['Date'], errors='coerce')
                if 'EntryDate' in df_trans.columns:
                    df_trans['Date'] = df_trans['Date'].fillna(pd.to_datetime(df_trans['EntryDate'], errors='coerce'))
                
                if active_ids:
                    df_trans = df_trans[df_trans['MaterialID'].apply(self.normalize_id).isin(active_ids)]
            
            # 2b. Prepare Daily Usage Records
            df_daily = pd.DataFrame()
            if hasattr(self, 'daily_usage_df') and not self.daily_usage_df.empty:
                df_daily = self.daily_usage_df.copy()
                # [FIX] Normalize column names to ensure reliable extraction of NDT_ fields
                df_daily.columns = [str(c).strip().replace(' ', '') for c in df_daily.columns]
                
                df_daily['Date'] = pd.to_datetime(df_daily['Date'], errors='coerce')
                # If Date is missing but EntryTime exists, use EntryTime
                if 'EntryTime' in df_daily.columns:
                    df_daily['Date'] = df_daily['Date'].fillna(pd.to_datetime(df_daily['EntryTime'], errors='coerce'))
                
            print(f"DEBUG: update_transaction_view - TRANS: {len(df_trans)}, DAILY: {len(df_daily)}")
            def _f(val):
                if val is None or pd.isna(val): return 0.0
                try: 
                    s = str(val).replace(',', '').strip()
                    return float(s) if s else 0.0
                except: return 0.0
            # [FIX] Do NOT filter by active_ids here; process everything and handle missing info later

            # 3. Create Harmonized Unified Data
            unified_rows = []
                
            # Add Transaction rows (Source: TRANS)
            if not df_trans.empty:
                for idx, row in df_trans.iterrows():
                    mat_id = self.normalize_id(row.get('MaterialID'))
                    mat_name = self.get_material_name_only(mat_id)
                    mat_info = self.get_material_info(mat_id)
                    
                    # [NEW] Filter: Only show consumables in In/Out history
                    if not self._is_consumable_material(mat_name, ''):
                        continue

                    # [NEW] Deduplication: Skip auto-generated transactions that are already covered by DAILY usage entries
                    # This removes the "50 vs -50" duplication while keeping the richer DAILY info.
                    note = str(row.get('Note', ''))
                    if "(자동 차감)" in note:
                        continue # Skip ALL auto-generated TRANS records as they are aggregated in the DAILY record

                    unified_rows.append({
                        'Date': row['Date'],
                        'Site': row.get('Site', ''),
                        'Type': row.get('Type', 'OUT'),
                        '품목명': mat_name,
                        'SN': mat_info.get('SN', ''),
                        '규격': mat_info.get('규격', ''),
                        'MaterialID': mat_id,
                        'Quantity': row.get('Quantity', 0),
                        'Warehouse': row.get('Warehouse', ''),
                        'User': row.get('User', ''),
                        'Note': row.get('Note', ''),
                        '차량번호': row.get('차량번호', ''),
                        'Source': 'TRANS',
                        'Fee': 0, # Transactions usually don't have inspection fees directly
                        'OrigIdx': idx
                    })


            # Add Daily Usage rows (Source: DAILY)
            if not df_daily.empty:
                for idx, row in df_daily.iterrows():
                    # [FLEXIBLE] Extract usage with multiple fallback names
                    u_val = _f(row.get('Usage', row.get('검사량', row.get('수량', row.get('Quantity', 0)))))
                    f_val = _f(row.get('FilmCount', row.get('매수', 0)))
                    
                    # If everything is 0, we still might want to show it if there's a note or site info
                    # but for now let's just make sure we capture NDT properly
                    ndt_vals_list = [
                        _f(row.get(f'NDT_{n}', row.get(n, 0))) 
                        for n in ['세척제', '침투제', '현상제', '백색페인트', '흑색자분', '형광자분', '형광침투제', '자분페인트']
                    ]
                    
                    # If any sign of life, keep the row
                    if u_val <= 0 and f_val <= 0 and not any(v > 0 for v in ndt_vals_list) and not str(row.get('Site', '')):
                        continue

                    mat_id = self.normalize_id(row.get('MaterialID'))
                    mat_name = self.get_material_name_only(mat_id)
                    mat_info = self.get_material_info(mat_id)
                    
                    # [NEW] Filter: Only show consumables in In/Out history (Exclude PAUT, etc.)
                    # If mat_name is empty (manual entry), check method or keywords in Note if needed, 
                    # but _is_consumable_material already handles method.
                    if not self._is_consumable_material(mat_name, row.get('검사방법', '')):
                        # Check if it has NDT chemicals (which ARE consumables)
                        if not any(v > 0 for v in ndt_vals_list):
                            continue
                    
                    # [FIX] Filter removed to show all site usage in history
                    
                    # Collect all worker names from User, User2...User10
                    workers = []
                    for i in range(1, 11):
                        u_col = 'User' if i == 1 else f'User{i}'
                        u_val_str = str(row.get(u_col, '')).strip()
                        if u_val_str and u_val_str not in ['nan', 'None', '']:
                            # Handle (주간) suffix if present
                            u_clean = re.sub(r'\(.*?\)\s*', '', u_val_str).strip()
                            if u_clean: workers.append(u_clean)
                    
                    worker_str = ", ".join(workers) if workers else str(row.get('User', ''))

                    # [FIX] Allow showing quantity for PT/MT/NDT items in history view
                    disp_qty = u_val

                    # [ROBUST] Try to parse ndt_data if it exists as a JSON string
                    ndt_json = {}
                    nj_raw = row.get('ndt_data', '')
                    if isinstance(nj_raw, str) and nj_raw.strip().startswith('{'):
                        try: ndt_json = json.loads(nj_raw)
                        except: pass
                    elif isinstance(nj_raw, dict):
                        ndt_json = nj_raw

                    unified_rows.append({
                        'Date': row['Date'],
                        'Site': row.get('Site', ''),
                        'Type': 'OUT',
                        '품목명': mat_name,
                        'SN': mat_info.get('SN', ''),
                        '규격': mat_info.get('규격', ''),
                        'MaterialID': mat_id,
                        'Quantity': disp_qty,
                        'NDT_세척제': _f(row.get('NDT_세척제', row.get('세척제', ndt_json.get('세척제', ndt_json.get('PT Cleaner', ndt_json.get('PTCleaner', 0)))))),
                        'NDT_침투제': _f(row.get('NDT_침투제', row.get('침투제', ndt_json.get('침투제', ndt_json.get('PT Penetrant', ndt_json.get('PTPenetrant', 0)))))),
                        'NDT_현상제': _f(row.get('NDT_현상제', row.get('현상제', ndt_json.get('현상제', ndt_json.get('PT Developer', ndt_json.get('PTDeveloper', 0)))))),
                        'NDT_백색페인트': _f(row.get('NDT_백색페인트', row.get('NDT_페인트', row.get('백색페인트', ndt_json.get('백색페인트', ndt_json.get('MT WHITE', ndt_json.get('MT7C-WHITE', 0))))))),
                        'NDT_흑색자분': row.get('NDT_흑색자분', row.get('NDT_자분', ndt_json.get('흑색자분', ndt_json.get('MT 7C-BLACK', 0)))),
                        'Warehouse': '',
                        'User': worker_str,
                        'Note': row.get('Note', ''),
                        '차량번호': row.get('차량번호', ''),
                        'Source': 'DAILY',
                        'Fee': row.get('검사비', 0.0), # Bring fee from daily usage
                        'OrigIdx': idx
                    })
            
            if not unified_rows:
                return
                
            df_unified = pd.DataFrame(unified_rows)
            
            # Apply Filters
            # Material Filter
            if hasattr(self, 'cb_trans_filter_mat'):
                selected_mat = str(self.cb_trans_filter_mat.get()).strip()
                if selected_mat and selected_mat != "전체":
                    # Check if MaterialID matches the display name in materials_df
                    matching_rows = []
                    for idx, row in df_unified.iterrows():
                        if self.get_material_display_name(row['MaterialID']) == selected_mat:
                            matching_rows.append(idx)
                    df_unified = df_unified.loc[matching_rows]
            
            # Site Filter
            if hasattr(self, 'cb_trans_filter_site'):
                selected_site = str(self.cb_trans_filter_site.get()).strip()
                if selected_site and selected_site != "전체":
                    df_unified = df_unified[df_unified['Site'].astype(str).str.contains(selected_site, na=False, case=False, regex=False)]
            
            # Vehicle Filter
            if hasattr(self, 'cb_trans_filter_vehicle'):
                selected_vehicle = str(self.cb_trans_filter_vehicle.get()).strip()
                if selected_vehicle and selected_vehicle != "전체":
                    df_unified = df_unified[df_unified['차량번호'].astype(str).str.contains(selected_vehicle, na=False, case=False, regex=False)]

            # [NEW] Update Filter Dropdowns with latest unified data unique values
            if hasattr(self, 'cb_trans_filter_site') and not getattr(self, '_site_filter_busy', False):
                self._site_filter_busy = True
                try:
                    s_current = self.cb_trans_filter_site.get()
                    s_vals = sorted(df_unified['Site'].dropna().astype(str).unique().tolist())
                    self.cb_trans_filter_site['values'] = ["전체"] + [v for v in s_vals if v.strip()]
                    if s_current: self.cb_trans_filter_site.set(s_current)
                    else: self.cb_trans_filter_site.set("전체")
                finally: self._site_filter_busy = False

            if hasattr(self, 'cb_trans_filter_mat') and not getattr(self, '_mat_filter_busy', False):
                self._mat_filter_busy = True
                try:
                    m_current = self.cb_trans_filter_mat.get()
                    # Get display names for all MaterialIDs in the unified set
                    unique_mat_ids = df_unified['MaterialID'].dropna().unique()
                    m_vals = sorted([self.get_material_display_name(mid) for mid in unique_mat_ids])
                    self.cb_trans_filter_mat['values'] = ["전체"] + [v for v in m_vals if v.strip()]
                    if m_current: self.cb_trans_filter_mat.set(m_current)
                    else: self.cb_trans_filter_mat.set("전체")
                finally: self._mat_filter_busy = False
                
            # [STABILITY] Ensure all standard columns are visible and reset any previous hiding
            all_inout_cols = ('날짜', '현장', '구분', '품목명', '수량', '세척제', '침투제', '현상제', '백색페인트', '흑색자분', 'SN', '규격', '창고', '담당자', '비고', '검사비')
            
            # Use saved configuration if available
            if hasattr(self, 'tab_config') and 'inout_visible_cols' in self.tab_config:
                self.inout_tree['displaycolumns'] = [c for c in self.tab_config['inout_visible_cols'] if c in all_inout_cols]
            else:
                self.inout_tree['displaycolumns'] = all_inout_cols

            # Sort and display
            df_sorted = df_unified.sort_values(by='Date', ascending=False).head(500)
            
            # [RESTORED] Column Order based on setup: 날짜, 구분, 품목명, SN, 규격, 수량, 현장, 창고, 담당자, 비고, 검사비
            for idx, row in df_sorted.iterrows():
                try:
                    # Robust date formatting with fallback for NaT
                    if pd.isna(row['Date']):
                        usage_date = "알 수 없음"
                    else:
                        usage_date = row['Date'].strftime('%Y-%m-%d %H:%M')
                    
                    # Manage Numeric Fee Formatting
                    f_val = row.get('Fee', 0.0)
                    fee_str = f"{f_val:,.0f}" if (isinstance(f_val, (int, float)) and f_val > 0) else ""

                    # Normalize Quantity for display: Use absolute value since 'Type' (IN/OUT) indicates direction
                    raw_qty = row.get('Quantity', row.get('수량', 0))
                    try:
                        clean_qty = abs(float(str(raw_qty).replace(',', '')))
                        qty_str = f"{clean_qty:.1f}" if clean_qty % 1 != 0 else f"{int(clean_qty)}"
                    except:
                        qty_str = self.clean_nan(raw_qty)

                    # NDT columns formatting
                    def _q_str(val):
                        try:
                            v = abs(float(str(val).replace(',', '')))
                            return f"{v:g}" if v > 0 else ""
                        except: return ""

                    self.inout_tree.insert('', tk.END, values=(
                        usage_date,
                        self.clean_nan(row.get('Site', '')),
                        self.clean_nan(row.get('Type', '-')),
                        self.clean_nan(row.get('품목명', '')),
                        qty_str,
                        _q_str(row.get('NDT_세척제', 0)),
                        _q_str(row.get('NDT_침투제', 0)),
                        _q_str(row.get('NDT_현상제', 0)),
                        _q_str(row.get('NDT_백색페인트', 0)),
                        _q_str(row.get('NDT_흑색자분', 0)),
                        self.clean_nan(row.get('SN', '')),
                        self.clean_nan(row.get('규격', '')),
                        self.clean_nan(row.get('Warehouse', '')),
                        self.clean_nan(row.get('User', '')),
                        self.clean_nan(row.get('Note', '')),
                        fee_str
                    ), tags=(str(row.get('OrigIdx', '')), row.get('Source', 'TRANS')))


                except Exception as e:
                    print(f"Row Display Error: {e}")
                    continue
                    
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Unified Transaction View Error: {e}\n{error_details}")

    def refresh_inout_history(self):
        """Refresh In/Out history and related filter source lists."""
        try:
            # Ensure filter candidates reflect latest data first
            self.update_material_combo()
            # Then redraw history and stock views
            self.update_transaction_view()
            self.update_stock_view()
        except Exception as e:
            messagebox.showerror("새로고침 오류", f"입출고 내역 새로고침 중 오류가 발생했습니다.\n{e}")

    def delete_transaction_entry(self):
        """Delete selected transaction from history and refresh"""
        selection = self.inout_tree.selection()
        if not selection:
            messagebox.showwarning("선택 오류", "삭제할 기록을 선택해주세요.")
            return
            
        if not messagebox.askyesno("삭제 확인", "선택한 기록을 삭제하시겠습니까?\n(삭제 시 재고 계산에 즉시 반영됩니다.)"):
            return
            
        for item in selection:
            tags = self.inout_tree.item(item, 'tags')
            if len(tags) >= 2:
                idx = int(tags[0])
                source = tags[1]
                
                if source == 'TRANS':
                    if idx in self.transactions_df.index:
                        self.transactions_df = self.transactions_df.drop(idx)
                elif source == 'DAILY':
                    if idx in self.daily_usage_df.index:
                        self.daily_usage_df = self.daily_usage_df.drop(idx)
                        # Optional: also delete from daily reports if needed
                        # But for now, removing from DailyUsage sheet is primary
        
        self.save_data()
        self.update_transaction_view()
        self.update_stock_view()
        messagebox.showinfo("완료", "거래 기록이 삭제되었습니다.")

    def on_material_selected(self, event=None):
        """Update the model listbox based on selected material"""
        selection = self.cb_material.get()
        if not selection:
            return
            
        # Clear existing models
        self.list_models.delete(0, tk.END)
        
        # Extract pure material name
        mat_name = selection
        if " - " in mat_name:
            mat_name = mat_name.split(" - ")[0]
        if " (SN: " in mat_name:
            mat_name = mat_name.split(" (SN: ")[0]
        
        pure_mat_name = mat_name
        
        # Find unique models for this material in materials_df
        if not self.materials_df.empty:
            relevant_mats = self.materials_df[self.materials_df['품목명'] == pure_mat_name]
            if not relevant_mats.empty:
                unique_models = relevant_mats['모델명'].dropna().unique()
                unique_models = sorted([str(m).strip() for m in unique_models if str(m).strip()])
                
                for model in unique_models:
                    self.list_models.insert(tk.END, model)
            # If no models found, add a placeholder
                if not unique_models:
                    self.list_models.insert(tk.END, "(등록된 모델명 없음)")

    def _on_trans_site_return(self, event):
        self.auto_save_to_list(event, self.cb_trans_site, self.sites, 'sites')
        self.cb_warehouse.focus_set()

    def _on_warehouse_return(self, event):
        self.auto_save_to_list(event, self.cb_warehouse, self.warehouses, 'warehouses')
        self.ent_user.focus_set()

    def _on_user_return(self, event):
        self.auto_save_to_list(event, self.ent_user, self.users, 'users')
        self.ent_note.focus_set()


    def format_worker_summary(self, workers):
        """Format a list of workers (or a joined string) into a compact summary string with a dropdown cue"""
        if not workers: return ""
        if isinstance(workers, str):
            if " | " in workers:
                names = [w.strip() for w in workers.split(" | ") if w.strip()]
            else:
                names = [w.strip() for w in workers.split(",") if w.strip()]
        else:
            names = [self.clean_nan(w) for w in workers if self.clean_nan(w)]
            
        unique_names = sorted(list(set(names)))
        if not unique_names: return ""
        
        return f"{unique_names[0]} [▼]"

    def show_worker_popup(self, event, tree):
        """Show a floating window (dropdown-style) with full record details for the clicked row"""
        item_id = tree.identify_row(event.y)
        column_id = tree.identify_column(event.x)
        if not item_id or not column_id: return
            
        col_name = tree.heading(column_id)['text']
        if col_name != '작업자': return
            
        tags = tree.item(item_id, 'tags')
        if not tags or 'total' in tags: return
            
        try:
            full_info = []
            if tree == getattr(self, 'daily_usage_tree', None):
                try:
                    df_idx = int(tags[0])
                    if df_idx not in self.daily_usage_df.index: return
                    entry = self.daily_usage_df.loc[df_idx]
                    
                    # 1. Basic Info
                    usage_date = self._safe_format_datetime(entry.get('Date', ''), '%Y-%m-%d')
                    full_info.append(f"[기본 기록 정보]")
                    full_info.append(f"• 날짜: {usage_date}")
                    full_info.append(f"• 현장: {self.clean_nan(entry.get('Site', ''))}")
                    
                    equip = self.clean_nan(entry.get('장비명', ''))
                    if equip: full_info.append(f"• 장비명: {equip}")
                    
                    method = self.clean_nan(entry.get('검사방법', ''))
                    if method: full_info.append(f"• 검사방법: {method}")
                    
                    amount = entry.get('검사량', 0)
                    if amount and float(amount) > 0: full_info.append(f"• 검사량: {amount}")
                    
                    
                    full_info.append(f"• 품목명: {self.get_material_display_name(entry.get('MaterialID'))}")
                    
                    # 2. Worker Info (All 10 slots)
                    full_info.append(f"\n[작업자 및 시간 정보]")
                    for i in range(1, 11):
                        u = self.clean_nan(entry.get('User' if i==1 else f'User{i}', ''))
                        if u:
                            t = self.clean_nan(entry.get('WorkTime' if i==1 else f'WorkTime{i}', ''))
                            o = self.clean_nan(entry.get('OT' if i==1 else f'OT{i}', ''))
                            full_info.append(f"• 작업자{i}: {u} | {t} | OT: {o}")
                    
                    # 3. Cost Info (Hide zeros)
                    cost_info = []
                    for k, label in [('단가', '단가'), ('출장비', '출장비'), ('일식', '일식'), ('검사비', '검사비')]:
                        raw_val = entry.get(k, 0)
                        try:
                            if pd.isna(raw_val) or str(raw_val).lower() == 'nan': continue
                            val_float = float(str(raw_val).replace(',', ''))
                            if val_float > 0:
                                cost_info.append(f"• {label}: {int(val_float):,}")
                        except: pass
                    if cost_info:
                        full_info.append(f"\n[비용 상세]")
                        full_info.extend(cost_info)
                    
                    # 4. RTK/NDT (Hide zeros)
                    rtk_found = []
                    rtk_cats = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타"]
                    for cat in rtk_cats:
                        val = entry.get(f'RTK_{cat}', 0)
                        if val and float(val) > 0:
                            rtk_found.append(f"  - {cat}: {val}")
                    if rtk_found:
                        full_info.append(f"\n[RT 결함 상세]")
                        full_info.extend(rtk_found)
                        
                    ndt_found = []
                    ndt_mats = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
                    for mat in ndt_mats:
                        val = entry.get(f'NDT_{mat}', 0)
                        if pd.isna(val): # Fallback
                            if mat == "흑색자분": val = entry.get('NDT_자분', 0)
                            elif mat == "백색페인트": val = entry.get('NDT_페인트', 0)
                        if val and float(val) > 0:
                            ndt_found.append(f"  - {mat}: {val}")
                    if ndt_found:
                        full_info.append(f"\n[NDT 자재 사용량]")
                        full_info.extend(ndt_found)
                    
                    note = entry.get('Note', '')
                    if note and str(note).strip(): full_info.append(f"\n• 비고: {note}")
                    
                    full_info.append(f"• 입력시간: {entry.get('EntryTime', '')}")
                    
                except Exception as e: 
                    print(f"Popup error: {e}")
                    return
            else:
                # Fallback for other trees (Monthly usage, etc.)
                vals = tree.item(item_id, 'values')
                cols = tree['columns']
                try:
                    w_idx = cols.index('작업자')
                    t_idx = cols.index('작업시간')
                    raw_names = vals[w_idx]
                    raw_times = vals[t_idx]
                    full_info.append(f"작업자: {raw_names}")
                    full_info.append(f"작업시간: {raw_times}")
                except: return

            if not full_info: return
            
            # Get cell bounding box for precise positioning
            bbox = tree.bbox(item_id, column_id)
            if not bbox: return
            
            x, y, w, h = bbox
            root_x = tree.winfo_rootx() + x
            root_y = tree.winfo_rooty() + y
            
            popup = tk.Toplevel(self.root)
            popup.overrideredirect(True)
            popup.attributes("-topmost", True)
            popup.configure(bg='white', highlightbackground="#0078d7", highlightthickness=2)
            
            popup_width = max(w, 550)
            popup_height = 450
            
            # 1. Horizontal positioning (Center relative to cell)
            final_x = root_x - (popup_width - w) // 2
            
            # 2. Vertical positioning (Show ABOVE the cell by default)
            final_y = root_y - popup_height - 5
            
            # [SMART CHECK] If the popup would go off-screen at the top, show it BELOW instead
            if final_y < 0:
                final_y = root_y + h + 5
            
            # Initial relative offsets from the cell root
            popup._rel_offset_x = final_x - root_x
            popup._rel_offset_y = final_y - root_y

            popup.geometry(f"{popup_width}x{popup_height}+{int(final_x)}+{int(final_y)}")
            
            # --- Custom Dragging Support ---
            def start_drag(event):
                popup._drag_data = {"x": event.x, "y": event.y}
            
            def do_drag(event):
                if hasattr(popup, '_drag_data'):
                    dx = event.x - popup._drag_data["x"]
                    dy = event.y - popup._drag_data["y"]
                    new_x = popup.winfo_x() + dx
                    new_y = popup.winfo_y() + dy
                    popup.geometry(f"+{new_x}+{new_y}")
                    
                    # Update relative offsets after manual drag so sync continues from new spot
                    try:
                        curr_bbox = tree.bbox(item_id, column_id)
                        if curr_bbox:
                            cx, cy, cw, ch = curr_bbox
                            # Important: Update relative to the CURRENT root position
                            popup._rel_offset_x = new_x - (tree.winfo_rootx() + cx)
                            popup._rel_offset_y = new_y - (tree.winfo_rooty() + cy)
                    except: pass
            
            popup.bind("<Button-1>", start_drag)
            popup.bind("<B1-Motion>", do_drag)
            
            content_frame = ttk.Frame(popup, padding=10)
            content_frame.pack(fill='both', expand=True)
            
            lb = tk.Listbox(content_frame, font=('Malgun Gothic', 10), 
                            bg='white', fg='#333333', relief='flat', 
                            borderwidth=0, highlightthickness=0)
            for info in full_info:
                lb.insert(tk.END, info)
            
            lb.pack(side='left', fill='both', expand=True)
            
            sb = ttk.Scrollbar(content_frame, orient="vertical", command=lb.yview)
            lb.configure(yscrollcommand=sb.set)
            sb.pack(side='right', fill='y')
            
            popup.focus_set()
            
            # --- Window Move Synchronization ---
            def reposition_popup(event=None):
                if not popup.winfo_exists(): return
                # Re-calculate position based on CURRENT root position + saved relative offset
                try:
                    curr_bbox = tree.bbox(item_id, column_id)
                    if not curr_bbox: return
                    cx, cy, cw, ch = curr_bbox
                    croot_x = tree.winfo_rootx() + cx
                    croot_y = tree.winfo_rooty() + cy
                    
                    # Apply relative offset to keep the sync accurate
                    new_x = croot_x + popup._rel_offset_x
                    new_y = croot_y + popup._rel_offset_y
                    
                    popup.geometry(f"+{int(new_x)}+{int(new_y)}")
                except: pass
                
            # Bind to root move/resize
            bind_id = self.root.bind("<Configure>", reposition_popup, add="+")
            
            def close_popup(e=None):
                if popup.winfo_exists():
                    try: self.root.unbind("<Configure>", bind_id)
                    except: pass
                    popup.destroy()
            
            popup.bind("<FocusOut>", close_popup)
            popup.bind("<Escape>", close_popup)
            popup.bind("<Double-1>", close_popup)
            
            # [FIX] Removed global self.root.bind("<Button-1>") which leaked and interfered with DateEntry.
            # FocusOut is sufficient for closing the popup when clicking elsewhere.
        except Exception as e:
            print(f"Final popup error: {e}")
    def clean_df_export(self, df):
        """Sanitize data for Excel: remove illegal characters, handle empty values, and protect formulas"""
        # 1. Pre-clean common empty representations
        df = df.replace(['nan', 'NaN', 'None'], "")
        
        # 2. Identify columns that have at least one meaningful value
        def is_really_empty(col):
            # A column is empty if all stripped values are "", NaN, or 0 sequences
            return df[col].astype(str).str.strip().replace(['nan', 'None', '', '0', '0.0', '0.00'], pd.NA).dropna().empty

        non_empty_cols = [col for col in df.columns if not is_really_empty(col)]
        
        # Always keep essential columns even if empty
        essential = ['날짜', '현장', '품목명', 'Date', 'Site']
        # Also keep standard RTK and NDT result columns
        rtk_cats = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타", "RTK총계"]
        ndt_materials = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
        essential.extend(rtk_cats)
        essential.extend([f'RTK_{c}' for c in rtk_cats])
        essential.extend(ndt_materials)
        essential.extend([f'NDT_{"".join(m.split())}' for m in ndt_materials])
        
        for col in essential:
            if col in df.columns and col not in non_empty_cols:
                non_empty_cols.append(col)
                
        # 3. Sanitize content and protect formulas
        def sanitize_and_escape(val):
            if pd.isna(val) or val == "" or val == 0 or val == 0.0:
                return ""
            
            # String conversion and cleaning
            s_val = str(val).strip()
            if not s_val or s_val.lower() in ['nan', 'none']:
                return ""

            # Remove illegal XML control characters (ASCII < 32 except tab, LF, CR)
            # These characters cause Excel to report file corruption.
            s_val = "".join(ch for ch in s_val if ord(ch) >= 32 or ch in "\t\n\r")
            
            # Escape potential formula characters (=, +, -, @)
            # IMPORTANT: Skip escaping if the value is a valid number (e.g., -5000)
            if s_val.startswith(('=', '+', '-', '@')):
                try:
                    # If it's a number, don't prefix with '
                    float(s_val)
                    return val
                except ValueError:
                    # It's a text string starting with a formula char
                    return f"'{s_val}"
            
            return s_val

        # Apply sanitization to all columns
        for col in df.columns:
            df[col] = df[col].apply(sanitize_and_escape)
        
        # Return only non-empty columns in original order
        final_cols = [c for c in df.columns if c in non_empty_cols]
        return df[final_cols]

    def save_df_to_excel_autofit(self, df, save_path, sheet_name='Sheet1'):
        """Save a DataFrame to Excel with automatic column width adjustment (AutoFit)"""
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            
            for idx, col in enumerate(df.columns):
                # Calculate max length of values in column + header
                # We handle Korean characters by assuming they take ~2 units of width
                def get_display_width(s):
                    width = 0
                    for char in str(s):
                        if ord(char) > 127: # Non-ASCII (Korean, etc.)
                            width += 2
                        else:
                            width += 1
                    return width

                series = df[col].astype(str)
                # Filter out empty strings/NAs for max calc
                lengths = series.apply(get_display_width)
                max_val_len = lengths.max() if not lengths.empty else 0
                header_len = get_display_width(col)
                
                # Final width with padding - Cap at a reasonable width to force wrapping
                final_width = min(max(max_val_len, header_len) + 2, 15)
                
                # Map index to column letter
                # column_letter property is available in openpyxl cells
                col_letter = worksheet.cell(row=1, column=idx+1).column_letter
                worksheet.column_dimensions[col_letter].width = final_width

            # --- Page Setup for Printing ---
            # Set to Portrait orientation (User request)
            worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
            
            # Enable Fit to Page width (Scale everything to fit horizontally)
            worksheet.sheet_properties.pageSetUpPr.fitToPage = True
            worksheet.page_setup.fitToWidth = 1
            worksheet.page_setup.fitToHeight = 0 # Automatic heights (multiple pages if long)
            
            # Enable Wrap Text for all cells to use vertical space instead of horizontal width
            from openpyxl.styles import Alignment
            wrap_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = wrap_alignment

            # Optional: Set small margins to maximize space (units are inches)
            worksheet.page_margins.left = 0.25
            worksheet.page_margins.right = 0.25
            worksheet.page_margins.top = 0.5
            worksheet.page_margins.bottom = 0.5


    
    def setup_import_tab(self):
        import_frame = ttk.LabelFrame(self.tab_import, text="데이터 관리")
        import_frame.pack(pady=20, padx=20, fill='both', expand=True)
        
        # Import Section
        ttk.Label(import_frame, text="엑셀 파일에서 자재 데이터 가져오기", font=('Arial', 11, 'bold')).pack(pady=10)
        ttk.Label(import_frame, text="형식: MaterialID, 회사코드, 관리품번, 품목명, 창고, 모델명, 규격, 품목군코드, 제조사, 제조국, 가격, 관리단위, 수량", 
                 wraplength=600).pack(pady=5)
        
        btn_import = ttk.Button(import_frame, text="엑셀 파일 가져오기", command=self.import_from_excel)
        btn_import.pack(pady=10)
        
        ttk.Separator(import_frame, orient='horizontal').pack(fill='x', pady=20)
        
        # Export Section
        ttk.Label(import_frame, text="현재 데이터 엑셀로 내보내기", font=('Arial', 11, 'bold')).pack(pady=10)
        
        btn_export_materials = ttk.Button(import_frame, text="자재 목록 내보내기", command=self.export_materials)
        btn_export_materials.pack(pady=5)
        
        btn_export_trans = ttk.Button(import_frame, text="거래 내역 내보내기", command=self.export_transactions)
        btn_export_trans.pack(pady=5)
        
        btn_export_all = ttk.Button(import_frame, text="전체 데이터 내보내기", command=self.export_all)
        btn_export_all.pack(pady=5)
    
    def import_from_excel(self):
        file_path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if not file_path:
            return
        
        try:
            # Load raw data
            raw_imported_df = pd.read_excel(file_path)
            
            # 1. Standardize column names (mapping) summerly
            col_map = {
                '품목명': ['품목명', '품명', 'Item Name', 'Name'],
                'SN': ['SN', 'SN번호', 'Serial Number', 'S/N'],
                '규격': ['규격', 'Specification', 'Spec'],
                '수량': ['수량', 'Initial Stock', 'Current Stock', 'Qty', 'Amount'],
                '모델명': ['모델명', 'Model', 'Machine'],
                '창고': ['창고', 'Warehouse', 'Location'],
                '가격': ['가격', 'Price', 'Unit Price'],
                '제조사': ['제조사', 'Manufacturer', 'Brand'],
                '공급업체': ['공급업체', '공급업자', 'Supplier', 'Vendor'],
                '관리단위': ['관리단위', 'Unit', 'UOM'],
                '재고하한': ['재고하한', '재주문 수준', 'Reorder Point', 'Min Stock']
            }
            
            # Simple column normalization
            working_df = pd.DataFrame()
            for std_name, aliases in col_map.items():
                for alias in aliases:
                    if alias in raw_imported_df.columns:
                        working_df[std_name] = raw_imported_df[alias]
                        break
                if std_name not in working_df.columns:
                    if std_name in ['품목명', '수량']: # Critical columns defaults
                        if std_name == '수량': working_df[std_name] = 0
                        else: 
                            messagebox.showerror("오류", f"필수 컬럼 '{std_name}'을(를) 찾을 수 없습니다.")
                            return
                    else:
                        working_df[std_name] = ""

            # Standardize values: clean strings, handle NaNs summerly
            working_df['품목명'] = working_df['품목명'].astype(str).str.strip()
            working_df['SN'] = working_df['SN'].astype(str).replace('nan', '').str.strip()
            working_df['규격'] = working_df['규격'].astype(str).replace('nan', '').str.strip()
            working_df['수량'] = pd.to_numeric(working_df['수량'], errors='coerce').fillna(0)
            working_df['가격'] = pd.to_numeric(working_df['가격'], errors='coerce').fillna(0)
            
            # [Added] Extract SN from Model name if necessary summerly
            def _extract_row_sn(row):
                m, s = self.extract_sn_from_model(row['모델명'], row['SN'])
                return pd.Series([m, s])
            
            working_df[['모델명', 'SN']] = working_df.apply(_extract_row_sn, axis=1)
            
            # Filter out rows without a name
            working_df = working_df[working_df['품목명'] != 'nan']
            working_df = working_df[working_df['품목명'] != '']
            
            # 2. Interior Deduplication: merge same items within the Excel file itself summerly
            # We group by Name, SN, and Spec.
            final_excel_df = working_df.groupby(['품목명', 'SN', '규격'], as_index=False).agg({
                '수량': 'sum',
                '가격': 'max',
                '모델명': 'first',
                '창고': 'first',
                '제조사': 'first',
                '공급업체': 'first',
                '관리단위': 'first',
                '재고하한': 'first'
            })
            
            # 3. Preparation: Match with existing inventory
            count_new = 0
            count_merged = 0
            count_overwritten = 0
            count_skipped = 0
            
            duplicate_indices = [] # (excel_row_idx, db_row_idx)
            new_items = []
            
            for idx, row in final_excel_df.iterrows():
                mat_name, sn, spec = row['품목명'], row['SN'], row['규격']
                
                existing_idx = -1
                if not self.materials_df.empty:
                    mask = (self.materials_df['품목명'].astype(str) == mat_name) & \
                           (self.materials_df['SN'].astype(str).replace('nan', '') == sn) & \
                           (self.materials_df['규격'].astype(str).replace('nan', '') == spec)
                    matches = self.materials_df.index[mask].tolist()
                    if matches:
                        existing_idx = matches[0]
                
                if existing_idx != -1:
                    duplicate_indices.append((idx, existing_idx))
                else:
                    new_items.append(idx)
            
            # 4. Conflict Resolution Choice Dialog summerly
            mode = "SKIP"
            if duplicate_indices:
                choices_dlg = tk.Toplevel(self.root)
                choices_dlg.title("중복 항목 처리 선택")
                choices_dlg.geometry("450x300")
                choices_dlg.transient(self.root)
                choices_dlg.grab_set()
                
                selection = tk.StringVar(value="CANCEL")
                
                ttk.Label(choices_dlg, text=f"기존 재고 목록과 일치하는 항목 {len(duplicate_indices)}건이 발견되었습니다.", 
                          font=('Malgun Gothic', 10, 'bold')).pack(pady=15)
                ttk.Label(choices_dlg, text="어떻게 처리할까요?").pack(pady=5)
                
                def set_choice(c):
                    selection.set(c)
                    choices_dlg.destroy()
                
                btn_frame = ttk.Frame(choices_dlg, padding=10)
                btn_frame.pack(fill='both', expand=True)
                
                ttk.Button(btn_frame, text="기존 재고에 수량 합치기 (추천)", 
                           command=lambda: set_choice("MERGE"), width=40).pack(pady=5)
                ttk.Button(btn_frame, text="기존 정보 덮어쓰기 (Excel 정보로 교체)", 
                           command=lambda: set_choice("OVERWRITE"), width=40).pack(pady=5)
                ttk.Button(btn_frame, text="중복 항목 건너뛰기 (신규만 추가)", 
                           command=lambda: set_choice("SKIP"), width=40).pack(pady=5)
                ttk.Button(btn_frame, text="가져오기 취소", 
                           command=lambda: set_choice("CANCEL"), width=40).pack(pady=15)
                
                self.root.wait_window(choices_dlg)
                mode = selection.get()
                if mode == "CANCEL":
                    return

            # 5. Process Import based on Mode summerly
            # Add new items first
            new_rows = []
            for idx in new_items:
                row_data = final_excel_df.iloc[idx].to_dict()
                new_mat_id = self.materials_df['MaterialID'].max() + 1 if not self.materials_df.empty else 1
                row_data['MaterialID'] = new_mat_id
                row_data['Active'] = 1
                new_rows.append(row_data)
                count_new += 1
            
            if new_rows:
                self.materials_df = pd.concat([self.materials_df, pd.DataFrame(new_rows)], ignore_index=True)
            
            # Handle duplicates
            if mode != "SKIP":
                for ex_idx, db_idx in duplicate_indices:
                    ex_row = final_excel_df.iloc[ex_idx]
                    
                    if mode == "MERGE":
                        # Add to existing quantity
                        current_qty = pd.to_numeric(self.materials_df.at[db_idx, '수량'], errors='coerce')
                        if pd.isna(current_qty): current_qty = 0
                        self.materials_df.at[db_idx, '수량'] = current_qty + ex_row['수량']
                        # Also update Manufacturer/Supplier if empty in DB but present in Excel
                        for field in ['제조사', '공급업체', '규격', '창고']:
                            if field in self.materials_df.columns and field in ex_row:
                                if not self.materials_df.at[db_idx, field] or str(self.materials_df.at[db_idx, field]) == 'nan':
                                    self.materials_df.at[db_idx, field] = ex_row[field]
                        count_merged += 1
                        
                    elif mode == "OVERWRITE":
                        # Full replace (except ID and Active status)
                        for col in final_excel_df.columns:
                            if col in self.materials_df.columns:
                                self.materials_df.at[db_idx, col] = ex_row[col]
                        count_overwritten += 1
            else:
                count_skipped = len(duplicate_indices)

            # 6. Finalize summerly
            self.save_data()
            self.update_material_combo()
            self.update_stock_view()
            self.update_registration_combos()
            
            summary = f"자재 가져오기가 완료되었습니다.\n\n"
            summary += f"• 신규 등록: {count_new}건\n"
            if mode == "MERGE":
                summary += f"• 재고량 합산: {count_merged}건\n"
            elif mode == "OVERWRITE":
                summary += f"• 정보 덮어쓰기: {count_overwritten}건\n"
            elif mode == "SKIP":
                summary += f"• 중복 건너뜀: {count_skipped}건\n"
            
            messagebox.showinfo("가져오기 완료", summary)
            
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messagebox.showerror("오류", f"파일을 가져오는데 실패했습니다: {e}")
            
        except Exception as e:
            messagebox.showerror("오류", f"파일을 가져오는데 실패했습니다: {e}")
    
    def export_materials(self):
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="Materials_Export.xlsx",
            title="자재 목록 저장",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if save_path:
            try:
                self.materials_df.to_excel(save_path, index=False)
                messagebox.showinfo("완료", "자재 목록이 저장되었습니다.")
            except Exception as e:
                messagebox.showerror("오류", f"저장 실패: {e}")
    
    def export_transactions(self):
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="Transactions_Export.xlsx",
            title="거래 내역 저장",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if save_path:
            try:
                self.transactions_df.to_excel(save_path, index=False)
                messagebox.showinfo("완료", "거래 내역이 저장되었습니다.")
            except Exception as e:
                messagebox.showerror("오류", f"저장 실패: {e}")
    
    def export_transaction_history(self):
        """Export the detailed In/Out history to Excel summerly"""
        history_data = []
        columns = [self.inout_tree.heading(col, 'text') for col in self.inout_tree['columns']]
        
        for item in self.inout_tree.get_children():
            values = self.inout_tree.item(item, 'values')
            row_data = {}
            for i, col in enumerate(columns):
                row_data[col] = values[i] if i < len(values) else ''
            history_data.append(row_data)
        
        if not history_data:
            messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
            return

    def _migrate_worktimes(self, input_list):
        """Helper to standardize work times with 익일 marker and sort logically"""
        default_worktimes = [
            "09:00~18:00", "09:00~19:00", "09:00~20:00", "09:00~21:00", 
            "09:00~22:00", "09:00~23:00", "09:00~24:00", "09:00~익일01:00",
            "09:00~익일02:00", "09:00~익일03:00", "18:00~익일02:00", "18:00~익일03:00"
        ]
        
        processed_times = set()
        for t in list(input_list) + default_worktimes:
            if not t or '~' not in t: continue
            
            clean_t = str(t).replace('익일', '').strip()
            try:
                start, end = clean_t.split('~')
                s_h = int(start.split(':')[0])
                e_h = int(end.split(':')[0])
                
                if e_h < s_h and '익일' not in str(t):
                    processed_times.add(f"{start}~익일{end}")
                else:
                    processed_times.add(str(t).strip())
            except:
                processed_times.add(str(t).strip())

        def time_sort_key(t):
            has_ikil = 1 if '익일' in t else 0
            clean = t.replace('익일', '')
            return (has_ikil, clean)

        return sorted(list(processed_times), key=time_sort_key)
        
        # Prepare filename
        today = datetime.datetime.now().strftime('%Y%m%d')
        filename = f"입출고내역_{today}.xlsx"
        
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=filename,
            title="입출고 내역 저장",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if save_path:
            try:
                history_df = pd.DataFrame(history_data)
                history_df = self.clean_df_export(history_df)
                self.save_df_to_excel_autofit(history_df, save_path, "입출고내역")
                messagebox.showinfo("완료", "입출고 내역이 저장되었습니다.")
            except Exception as e:
                messagebox.showerror("오류", f"저장 실패: {e}")
    
    def export_all(self):
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="Complete_Export.xlsx",
            title="전체 데이터 저장",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if save_path:
            try:
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    self.materials_df.to_excel(writer, sheet_name='Materials', index=False)
                    self.transactions_df.to_excel(writer, sheet_name='Transactions', index=False)
                messagebox.showinfo("완료", "전체 데이터가 저장되었습니다.")
            except Exception as e:
                messagebox.showerror("오류", f"저장 실패: {e}")
    

    def view_monthly_usage(self):
        """Display monthly usage by item name in the treeview"""
        # Clear current view
        for item in self.usage_tree.get_children():
            self.usage_tree.delete(item)
        
        year = int(self.cb_year.get())
        month = int(self.cb_month.get())
        
        # Filter transactions for the selected month
        month_mask = (self.transactions_df['Date'].dt.year == year) & \
                     (self.transactions_df['Date'].dt.month == month) & \
                     (self.transactions_df['Type'] == 'OUT')
        monthly_trans = self.transactions_df[month_mask]
        
        # Filter transactions for cumulative (from start of year to selected month)
        cumulative_mask = (self.transactions_df['Date'].dt.year == year) & \
                          (self.transactions_df['Date'].dt.month <= month) & \
                          (self.transactions_df['Type'] == 'OUT')
        cumulative_trans = self.transactions_df[cumulative_mask]
        
        # Build usage data for each material
        usage_data = []
        for _, mat in self.materials_df.iterrows():
            mat_id = mat['MaterialID']
            
            # Calculate monthly usage
            month_usage = monthly_trans[monthly_trans['MaterialID'] == mat_id]['Quantity'].sum()
            
            # Calculate cumulative usage (year-to-date)
            cumulative_usage = cumulative_trans[cumulative_trans['MaterialID'] == mat_id]['Quantity'].sum()
            
            # Only show items with usage
            if month_usage > 0 or cumulative_usage > 0:
                usage_data.append({
                    '품목명': mat.get('품목명', ''),  # Using 재고현황's 품목명 field
                    '관리품번': mat.get('관리품번', ''),
                    '규격': mat.get('규격', ''),
                    '단위': mat.get('관리단위', 'EA'),
                    '월사용량': month_usage,
                    '누계사용량': cumulative_usage
                })
        
        # Sort by item name
        usage_data.sort(key=lambda x: x['품목명'])
        
        # Display in treeview
        for data in usage_data:
            self.usage_tree.insert('', tk.END, values=(
                data['품목명'],
                data['관리품번'],
                data['규격'],
                data['단위'],
                f"{data['월사용량']:.1f}",
                f"{data['누계사용량']:.1f}"
            ))
        
        # Show message if no data
        if not usage_data:
            messagebox.showinfo("알림", f"{year}년 {month}월에 사용 내역이 없습니다.")

    def generate_yearly_report(self):
        year = int(self.cb_year.get())
        # Filter transactions for the year
        mask = (self.transactions_df['Date'].dt.year == year) & (self.transactions_df['Type'] == 'OUT')
        yearly_trans = self.transactions_df[mask]
        
        report = []
        for _, mat in self.materials_df.iterrows():
            mat_id = mat['MaterialID']
            
            # Calculate monthly usage for each month
            row_data = {
                '설비코드': mat.get('관리품번', ''),
                '자재명': mat.get('품목명', ''),  # Using 재고현황's 품목명 field
                '분류': mat.get('품목군코드', ''),
                '규격': mat.get('규격', ''),
                '단위': mat.get('관리단위', ''),
                '제조사': mat.get('제조사', '')
            }
            
            # Add monthly columns (1월 ~ 12월)
            monthly_values = []
            for month in range(1, 13):
                month_mask = (yearly_trans['MaterialID'] == mat_id) & \
                            (yearly_trans['Date'].dt.month == month)
                month_usage = yearly_trans[month_mask]['Quantity'].sum()
                row_data[f'{month}월'] = month_usage
                monthly_values.append(month_usage)
            
            # Calculate totals
            total = sum(monthly_values)
            row_data['합계'] = total
            
            # Calculate cumulative total
            cumulative = 0
            for i, val in enumerate(monthly_values, 1):
                cumulative += val
                if i == 12:  # Only show final cumulative at the end
                    row_data['누계'] = cumulative
            
            report.append(row_data)
            
        report_df = pd.DataFrame(report)
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                 initialfile=f"Yearly_Usage_{year}.xlsx",
                                                 title="보고서 저장")
        if save_path:
            report_df.to_excel(save_path, index=False)
            messagebox.showinfo("완료", f"{year}년 보고서가 저장되었습니다.")

    def generate_monthly_report(self):
        year = int(self.cb_year.get())
        month = int(self.cb_month.get())
        
        mask = (self.transactions_df['Date'].dt.year == year) & \
               (self.transactions_df['Date'].dt.month == month) & \
               (self.transactions_df['Type'] == 'OUT')
        monthly_trans = self.transactions_df[mask]
        
        report = []
        for _, mat in self.materials_df.iterrows():
            mat_id = mat['MaterialID']
            total_usage = monthly_trans[monthly_trans['MaterialID'] == mat_id]['Quantity'].sum()
            report.append({
                '설비코드': mat.get('관리품번', ''),
                '자재명': mat.get('품목명', ''),  # Using 재고현황's 품목명 field
                '분류': mat.get('품목군코드', ''),
                '규격': mat.get('규격', ''),
                '단위': mat.get('관리단위', ''),
                '제조사': mat.get('제조사', ''),
                '월간 총 사용량': total_usage
            })
            
        report_df = pd.DataFrame(report)
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                 initialfile=f"Monthly_Usage_{year}_{month}.xlsx",
                                                 title="보고서 저장")
        if save_path:
            report_df.to_excel(save_path, index=False)
            messagebox.showinfo("완료", f"{year}년 {month}월 보고서가 저장되었습니다.")

    def setup_monthly_usage_tab(self):
        """Setup the monthly usage aggregation tab (auto-aggregated from daily usage)"""
        # Display frame for aggregated monthly data
        display_frame = ttk.LabelFrame(self.tab_monthly_usage, text="월별 사용량 집계 (현장별 데이터 자동 집계)")
        display_frame.pack(expand=True, fill='both', padx=10, pady=10)
        
        # Filter controls
        filter_frame = ttk.Frame(display_frame)
        filter_frame.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(filter_frame, text="연도:").pack(side='left', padx=5)
        current_year = datetime.datetime.now().year
        year_values = ['전체'] + [str(y) for y in range(max(2024, current_year-1), current_year + 3)]
        self.cb_filter_year = ttk.Combobox(filter_frame, values=year_values, width=10)
        self.cb_filter_year.pack(side='left', padx=5)
        self.cb_filter_year.set('전체')
        
        ttk.Label(filter_frame, text="월:").pack(side='left', padx=5)
        current_month = datetime.datetime.now().month
        self.cb_filter_month = ttk.Combobox(filter_frame, values=['전체'] + [str(m) for m in range(1, 13)], width=10)
        self.cb_filter_month.pack(side='left', padx=5)
        self.cb_filter_month.set('전체')
        
        ttk.Label(filter_frame, text="현장:").pack(side='left', padx=5)
        self.cb_filter_site_monthly = ttk.Combobox(filter_frame, width=15)
        self.cb_filter_site_monthly.pack(side='left', padx=5)
        self.cb_filter_site_monthly.set('전체')
        
        btn_site_mgr_monthly = tk.Button(filter_frame, text="⚙", font=('Arial', 7), bd=0, bg=self.theme_bg, fg='gray',
                                       command=lambda: self.open_list_management_dialog('sites', target_cb=self.cb_filter_site_monthly))
        btn_site_mgr_monthly.place(in_=self.cb_filter_site_monthly, relx=1.0, x=-18, rely=0.5, anchor='e', width=16, height=16)
        
        ttk.Label(filter_frame, text="품목명:").pack(side='left', padx=5)
        self.cb_filter_material_monthly = ttk.Combobox(filter_frame, width=25)
        self.cb_filter_material_monthly.pack(side='left', padx=5)
        self.cb_filter_material_monthly.set('전체')
        
        btn_filter = ttk.Button(filter_frame, text="조회", command=self.update_monthly_usage_view)
        btn_filter.pack(side='left', padx=10)
        
        btn_export = ttk.Button(filter_frame, text="엑셀 내보내기", command=self.export_monthly_usage_history)
        btn_export.pack(side='left', padx=5)

        btn_popout = ttk.Button(filter_frame, text="🔍 팝업창으로 열기", command=self.open_detached_monthly_usage_view)
        btn_popout.pack(side='left', padx=5)
        
        # Use a PanedWindow to allow resizing between main tree and summaries
        self.monthly_paned = ttk.PanedWindow(display_frame, orient="vertical")
        self.monthly_paned.pack(expand=True, fill='both', padx=5, pady=5)
        
        # 1. Top pane: Main Monthly Usage Tree
        tree_frame = ttk.Frame(self.monthly_paned)
        self.monthly_paned.add(tree_frame, weight=3) # Give main tree more weight
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        # Treeview with columns including worker, work time, and OT fields
        # Added '(Full작업자)' for full list backup during Excel export
        columns = ('연도', '월', '현장', '작업자', '작업시간', 'OT시간', 'OT금액', 'OT1', 'OT2', 'OT3', 'OT4', 'OT5', 'OT6', 'OT7', 'OT8', 'OT9', 'OT10', 
                   '수량', '단가', '출장비', '일식', '검사비', '품목명', '센터미스', '농도', '마킹미스', '필름마크', 
                   '취급부주의', '고객불만', '기타', 'RTK총계', '형광자분', '흑색자분', '백색페인트', '침투제', '세척제', '현상제', '형광침투제', '비고', '(Full작업자)')
        self.monthly_usage_tree = ttk.Treeview(tree_frame, columns=columns, show='headings',
                                               yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Hide the (Full작업자) column from display
        self.monthly_usage_tree['displaycolumns'] = [c for c in columns if c != '(Full작업자)']
        
        vsb.config(command=self.monthly_usage_tree.yview)
        hsb.config(command=self.monthly_usage_tree.xview)
        
        # Column configuration with added OT columns
        col_widths = {
            '연도': 90, '월': 70, '현장': 140, '작업자': 100, '작업시간': 100,
            'OT시간': 100, 'OT금액': 110,
            'OT1': 100, 'OT2': 100, 'OT3': 100, 'OT4': 100, 'OT5': 100,
            'OT6': 100, 'OT7': 100, 'OT8': 100, 'OT9': 100, 'OT10': 100,
            '수량': 110, '단가': 110, '출장비': 110, '일식': 110, '검사비': 110,
            '품목명': 220, '수량': 110, '센터미스': 80, '농도': 80, '마킹미스': 80,
            '필름마크': 80, '취급부주의': 80, '고객불만': 80, '기타': 80, 'RTK총계': 80,
            '형광자분': 90, '흑색자분': 90, '백색페인트': 90, '침투제': 90, '세척제': 90,
            '현상제': 90, '형광침투제': 90, '비고': 220
        }
        
        for col in columns:
            self.monthly_usage_tree.heading(col, text=col, command=lambda c=col: self.treeview_sort_column(self.monthly_usage_tree, c, False))
            width = col_widths.get(col, 100)
            self.monthly_usage_tree.column(col, width=width, minwidth=20, stretch=False, anchor='center')
        
        # [NEW] Enable column reordering via drag & drop
        self.enable_tree_column_drag(self.monthly_usage_tree)
        
        
        # Grid layout
        self.monthly_usage_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # [NEW] Auto-save column widths when user resizes columns
        self.monthly_usage_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())
        
        # [NEW] Bind selection to update site/worker summaries
        self.monthly_usage_tree.bind('<<TreeviewSelect>>', self.on_monthly_usage_select)

        # 2. Middle pane: Site Summary
        site_frame = ttk.LabelFrame(self.monthly_paned, text="현장별 누계")
        self.monthly_paned.add(site_frame, weight=1)
        
        self.monthly_usage_tree.bind("<Button-1>", lambda e: self.show_worker_popup(e, self.monthly_usage_tree), add="+")
        
        site_cols = ('현장', '검사방법', '품목명', '수량', '검사비', '출장비', '형광자분', '흑색자분', '백색페인트', 
                     '침투제', '세척제', '현상제', '형광침투제', '센터미스', '농도', '마킹미스', '필름마크', '취급부주의', '고객불만', '기타', 'RTK총계')
        self.site_summary_tree = ttk.Treeview(site_frame, columns=site_cols, show='headings')
        for col in site_cols:
            self.site_summary_tree.heading(col, text=col)
            # Adjust widths based on content
            if col in ['현장', '검사방법', '품목명']: width = 120
            elif col in ['검사비', '출장비']: width = 100
            else: width = 80
            self.site_summary_tree.column(col, width=width, anchor='center', stretch=False)
        
        # [NEW] Auto-save site summary column widths
        self.site_summary_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())
        self.enable_tree_column_drag(self.site_summary_tree, context_menu_handler=lambda e: self._show_generic_tree_heading_context_menu(e, self.site_summary_tree))
        
        site_vsb = ttk.Scrollbar(site_frame, orient="vertical", command=self.site_summary_tree.yview)
        self.site_summary_tree.configure(yscrollcommand=site_vsb.set)
        
        self.site_summary_tree.pack(side='left', expand=True, fill='both')
        site_vsb.pack(side='right', fill='y')
        
        # 3. Bottom pane: Worker Summary
        worker_frame = ttk.LabelFrame(self.monthly_paned, text="작업자별 누계")
        self.monthly_paned.add(worker_frame, weight=1)
        
        worker_cols = ('작업자', '총공수', '연장(시간)', '야간(시간)', '휴일(시간)', '총OT(시간)', '연장(금액)', '야간(금액)', '휴일(금액)', '총OT(금액)')
        self.worker_summary_tree = ttk.Treeview(worker_frame, columns=worker_cols, show='headings')
        
        # Set column widths for worker summary
        worker_widths = {
            '작업자': 100, '총공수': 70, '연장(시간)': 70, '야간(시간)': 70, '휴일(시간)': 70, 
            '총OT(시간)': 80, '연장(금액)': 90, '야간(금액)': 90, '휴일(금액)': 90, '총OT(금액)': 100
        }
        for col in worker_cols:
            self.worker_summary_tree.heading(col, text=col)
            width = worker_widths.get(col, 80)
            # Disable stretching
            self.worker_summary_tree.column(col, width=width, anchor='center', stretch=False)
        
        # [NEW] Auto-save worker summary column widths
        self.worker_summary_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())
        self.enable_tree_column_drag(self.worker_summary_tree, context_menu_handler=lambda e: self._show_generic_tree_heading_context_menu(e, self.worker_summary_tree))
            
        worker_vsb = ttk.Scrollbar(worker_frame, orient="vertical", command=self.worker_summary_tree.yview)
        self.worker_summary_tree.configure(yscrollcommand=worker_vsb.set)
        
        self.worker_summary_tree.pack(side='left', expand=True, fill='both')
        worker_vsb.pack(side='right', fill='y')

        # [NEW] 공사탭 특별근무 자동 입력 버튼
        btn_apply_frame = ttk.Frame(worker_frame)
        btn_apply_frame.pack(fill='x', pady=3)
        ttk.Button(btn_apply_frame,
                   text="📋 공사탭 특별근무에 적용",
                   command=self.apply_worker_shift_hours_to_budget).pack(side='right', padx=5)

        # Initial view update
        self.update_monthly_usage_view()
    
    def update_monthly_usage_view(self):
        """Update the monthly usage treeview with aggregated data from daily usage"""
        # Clear current views
        for item in self.monthly_usage_tree.get_children():
            self.monthly_usage_tree.delete(item)
        for item in self.site_summary_tree.get_children():
            self.site_summary_tree.delete(item)
        for item in self.worker_summary_tree.get_children():
            self.worker_summary_tree.delete(item)
            
        # [NEW] Clear detached monthly views if open
        if 'monthly' in self.detached_windows:
            p_tree = self.detached_windows['monthly']['tree']
            for item in p_tree.get_children():
                p_tree.delete(item)
        
        # Get filter values
        filter_year = self.cb_filter_year.get()
        filter_month = self.cb_filter_month.get()
        filter_site = self.cb_filter_site_monthly.get() if hasattr(self, 'cb_filter_site_monthly') else '전체'
        filter_material = self.cb_filter_material_monthly.get() if hasattr(self, 'cb_filter_material_monthly') else '전체'
        
        # Return if daily usage data is empty
        if self.daily_usage_df.empty:
            return
        
        # Create a copy of daily usage data and extract year/month from Date column
        df = self.daily_usage_df.copy()
        # [CRITICAL] Normalize columns to ensure detection (matched_pairs) matches data lookups
        df.columns = [str(c).strip().replace(' ', '') for c in df.columns]
        
        # Normalize column names - remove ALL types of whitespace using regex
        import re
        df.columns = [re.sub(r'\s+', '', str(c)) for c in df.columns]
        
        df['Year'] = pd.to_datetime(df['Date'], errors='coerce').dt.year
        df['Month'] = pd.to_datetime(df['Date'], errors='coerce').dt.month
        # 날짜 파싱 실패 행 제거
        df = df.dropna(subset=['Year', 'Month'])
        df['Year'] = df['Year'].astype(int)
        df['Month'] = df['Month'].astype(int)
        
        # [ROBUST] Normalize Site names to prevent duplicate grouping/mismatches (e.g., merging "Site-A" and "Site - A")
        if 'Site' in df.columns:
            df['Site'] = df['Site'].astype(str).apply(self.normalize_site_name)
        
        # Apply filters
        if filter_year != '전체':
            df = df[df['Year'] == int(filter_year)]
        
        if filter_month != '전체':
            df = df[df['Month'] == int(filter_month)]
        
        if filter_site != '전체':
            df = df[df['Site'] == filter_site]
            
        if filter_material != '전체':
            # Get matching MaterialIDs for the selected item name
            matching_ids = self.materials_df[self.materials_df['품목명'] == filter_material]['MaterialID'].tolist()
            if matching_ids:
                df = df[df['MaterialID'].isin(matching_ids)]
            else:
                # If no material found in master list, clear and return (though this shouldn't happen with sync)
                return
        
        # [NEW] Return early if filtered df is empty to avoid ValueError during assignment
        if df.empty:
            return
        # Populate site filter options from data
        if hasattr(self, 'cb_filter_site_monthly'):
            # [ROBUST] Use same collection logic as refresh_inquiry_filters to exclude hidden sites and maintain sync
            raw_sites = set()
            if not self.daily_usage_df.empty and 'Site' in self.daily_usage_df.columns:
                raw_sites.update(self.daily_usage_df['Site'].dropna().astype(str).apply(self.normalize_site_name).tolist())
            if hasattr(self, 'sites'):
                for s in self.sites:
                    norm = self.normalize_site_name(s)
                    if norm: raw_sites.add(norm)
            if hasattr(self, 'budget_df') and not self.budget_df.empty and 'Site' in self.budget_df.columns:
                raw_sites.update(self.budget_df['Site'].dropna().astype(str).apply(self.normalize_site_name).tolist())
            
            unique_sites = ['전체'] + sorted([s for s in raw_sites 
                                             if s and str(s).lower() != 'nan'
                                            and s not in getattr(self, 'hidden_sites', [])])
            
            self.cb_filter_site_monthly['values'] = unique_sites
            if not self.cb_filter_site_monthly.get():
                self.cb_filter_site_monthly.set('전체')
        
        # Populate material filter options from data
        if hasattr(self, 'cb_filter_material_monthly'):
            # Get unique material names from materials_df based on MaterialIDs in daily_usage_df
            # Note: MaterialID column name itself might have spaces in self.daily_usage_df
            m_id_col = 'MaterialID' if 'MaterialID' in self.daily_usage_df.columns else 'MaterialID'
            unique_mat_ids = self.daily_usage_df[m_id_col].dropna().unique()
            material_names = []
            for mat_id in unique_mat_ids:
                # [ROBUST] Use same lookup logic for filters
                try:
                    m_id_f = float(mat_id)
                    matches = self.materials_df[pd.to_numeric(self.materials_df['MaterialID'], errors='coerce') == m_id_f]
                    if not matches.empty:
                        material_names.append(str(matches.iloc[0]['품목명']))
                    else:
                        material_names.append(f"ID: {mat_id}")
                except:
                    matches = self.materials_df[self.materials_df['MaterialID'].astype(str) == str(mat_id)]
                    if not matches.empty:
                        material_names.append(str(matches.iloc[0]['품목명']))
                    else:
                        material_names.append(f"ID: {mat_id}")
            unique_materials = ['전체'] + sorted(set(material_names))
            self.cb_filter_material_monthly['values'] = unique_materials
            if not self.cb_filter_material_monthly.get():
                self.cb_filter_material_monthly.set('전체')

        # [NEW] Sync detached filters to main window state
        if 'monthly' in self.detached_windows:
            p_filters = self.detached_windows['monthly'].get('filters', {})
            if p_filters:
                p_filters['year'].set(self.cb_filter_year.get())
                p_filters['month'].set(self.cb_filter_month.get())
                p_filters['site'].set(self.cb_filter_site_monthly.get())
                p_filters['mat'].set(self.cb_filter_material_monthly.get())
                
                # Also sync dropdown values for Site/Material
                p_filters['site']['values'] = self.cb_filter_site_monthly['values']
                p_filters['mat']['values'] = self.cb_filter_material_monthly['values']
        
        # Prepare aggregation dictionary for all numeric fields
        agg_dict = {'Usage': 'sum'}
        
        # Helper for joining workers - clean internal spaces to avoid "11시간" vs "11 시간" mismatch
        def join_unique_non_empty(series):
            # Strip outer spaces and compress internal spaces for consistency
            vals = [" ".join(str(v).split()) for v in series if pd.notna(v) and str(v).strip()]
            return " | ".join(sorted(set(vals)))
        
        # Also define a sum helper that handles potential type issues and commas
        def safe_sum(series):
            def to_f(v):
                if pd.isna(v) or str(v).lower() in ('nan', 'none', ''): return 0.0
                s = str(v).strip().lower()
                try: 
                    # Remove non-numeric markers (comma, unit, etc)
                    clean_s = re.sub(r'[^0-9\.\-]', '', s)
                    return float(clean_s) if clean_s else 0.0
                except: return 0.0
            return series.apply(to_f).sum()

        # [NEW] Hyper-Robust Column Detection for Workers/WorkTime/OT
        # Find all columns using exact name matching (not regex with optional suffix)
        def find_paired_cols(cols):
            pairs = []
            col_set = set(cols)
            for i in range(1, 11):
                # i=1: 'User', i=2: 'User2', etc.
                u_name = 'User' if i == 1 else f'User{i}'
                w_name = 'WorkTime' if i == 1 else f'WorkTime{i}'
                o_name = 'OT' if i == 1 else f'OT{i}'
                
                u_col = u_name if u_name in col_set else None
                w_col = w_name if w_name in col_set else None
                o_col = o_name if o_name in col_set else None
                
                if u_col: pairs.append((u_col, w_col, o_col))
            return pairs

        matched_pairs = find_paired_cols(df.columns)
        for u_c, w_c, o_c in matched_pairs:
            agg_dict[u_c] = join_unique_non_empty
            if w_c: agg_dict[w_c] = join_unique_non_empty
            if o_c: agg_dict[o_c] = join_unique_non_empty
        
        # Removed FilmCount aggregation as it is now integrated into Usage
        
        # Add RTK categories
        rtk_categories = ['RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크', 'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타']
        for cat in rtk_categories:
            if cat in df.columns:
                agg_dict[cat] = safe_sum
        
        # Add NDT materials and cost fields
        other_agg_cols = ['NDT_형광자분', 'NDT_자분', 'NDT_흑색자분', 'NDT_페인트', 'NDT_백색페인트', 'NDT_침투제', 'NDT_세척제', 'NDT_현상제', 'NDT_형광', 'NDT_형광침투제',
                          '검사량', '단가', '출장비', '일식', '검사비']
        for col in other_agg_cols:
            if col in df.columns:
                agg_dict[col] = safe_sum

        # [NEW] Pre-aggregation Deduping to ensure parity with Site Tab
        seen_m_times = set()
        seen_m_contents = set()
        
        # Robust mapping for worker columns in Monthly Tab
        matched_pairs = find_paired_cols(df.columns)

        def sync_dedup_and_calc_ot(row):
            # 1. Calculate Activity-based OT (Max of all workers in this row)
            h_max = 0.0
            a_sum = 0
            w_count = 0
            raw_workers = []
            
            for u_c, w_c, o_c in matched_pairs:
                u_v = self.clean_nan(row.get(u_c, ''))
                if u_v:
                    raw_workers.append(u_v)
                    w_count += 1
                
                if o_c:
                    ots = str(row.get(o_c, '')).strip()
                    if ots and ots not in ('nan', '0.0', '0'):
                        try:
                            # Use same parsing logic as Site Tab
                            if '(' in ots and '원)' in ots:
                                h_p = float(ots.split('시간')[0])
                                amt_str = _re.sub(r'[^0-9]', '', ots.split('(')[1].split('원')[0])
                                a_p = int(amt_str) if amt_str else 0
                            elif ots.replace(',', '').isdigit():
                                a_p = int(ots.replace(',', ''))
                                wt_v = str(row.get(w_c, '')).strip() if w_c else ''
                                h_p, _ = self._calculate_ot_from_worktime(wt_v, pd.to_datetime(row.get('Date', pd.Timestamp.now())))
                            else:
                                a_p = self.calculate_ot_amount(ots)
                                h_p = self._parse_ot_hours(ots)
                            
                            h_max = max(h_max, h_p)
                            a_sum += a_p
                        except: pass

            # 2. Deduping Key (Date, Site, WorkTime, Material)
            # [REFINED] Exclude workers from the key to correctly catch records split across rows.
            n_date = self._safe_format_datetime(row.get('Date', ''), '%Y-%m-%d')
            n_site = str(row.get('Site', '')).strip()
            c_worktime = str(row.get('WorkTime', '')).strip()
            n_mat = str(row.get('MaterialID', ''))
            
            content_key = (n_date, n_site, c_worktime, n_mat)
            
            e_t_raw = row.get('EntryTime', '')
            try:
                if isinstance(e_t_raw, (pd.Timestamp, datetime.datetime)):
                    t_key = e_t_raw.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    t_key = str(e_t_raw).split('.')[0].strip() if e_t_raw else ""
            except: t_key = ""
            
            is_dup = (t_key and t_key in seen_m_times) or (content_key in seen_m_contents)
            
            if not is_dup:
                if t_key: seen_m_times.add(t_key)
                seen_m_contents.add(content_key)
                
                # Primary row: return full values
                return pd.Series([h_max, a_sum, w_count, False])
            else:
                # Duplicate row: Zero out quantitative impact for aggregation
                return pd.Series([0.0, 0, 0, True])

        # Apply calculation and marking
        calc_results = df.apply(sync_dedup_and_calc_ot, axis=1)
        df[['OT시간', 'OT금액', 'WorkerCount', '_is_m_dup']] = calc_results
        
        # Zero out other quantitative fields for duplicate rows before aggregation
        # [FIX] Do NOT zero out 'Usage' and '검사량' during view-level deduping.
        # Database-level splitting already zeros them for valid splits. 
        # View-level zeroing was causing data loss if the quantity was on a secondary row.
        q_fields = ['단가', '출장비', '일식', '검사비', 'OT시간', 'OT금액']
        rtk_fields = [f'RTK_{c}' for c in ['센터미스', '농도', '마킹미스', '필름마크', '취급부주의', '고객불만', '기타']]
        ndt_fields = ['NDT_형광자분', 'NDT_자분', 'NDT_흑색자분', 'NDT_페인트', 'NDT_백색페인트', 'NDT_침투제', 'NDT_세척제', 'NDT_현상제', 'NDT_형광', 'NDT_형광침투제']
        
        dup_mask = df['_is_m_dup'] == True
        for f in q_fields + rtk_fields + ndt_fields:
            if f in df.columns:
                df.loc[dup_mask, f] = 0.0

        # Group by Year, Month, Site, MaterialID and aggregate
        agg_dict['OT시간'] = 'sum'
        agg_dict['OT금액'] = 'sum'
        agg_dict['WorkerCount'] = 'sum'
        grouped = df.groupby(['Year', 'Month', 'Site', 'MaterialID'], dropna=False).agg(agg_dict).reset_index()
        
        self.monthly_usage_tree.bind("<Button-1>", lambda e: self.show_worker_popup(e, self.monthly_usage_tree), add="+")
        
        # Store df for selection handling
        self.current_monthly_df = df
        
        # Initialize totals for cumulative sum
        total_worker_count = 0.0
        total_ot_hours = 0.0
        total_ot_amount = 0.0
        total_test_amount = 0.0
        total_unit_price = 0.0
        total_travel_cost = 0.0
        total_meal_cost = 0.0
        total_test_fee = 0.0
        total_rtk_center = 0.0
        total_rtk_density = 0.0
        total_rtk_marking = 0.0
        total_rtk_film = 0.0
        total_rtk_handling = 0.0
        total_rtk_customer = 0.0
        total_rtk_other = 0.0
        total_ndt_fluorescent_mag = 0.0
        total_ndt_magnet = 0.0
        total_ndt_paint = 0.0
        total_ndt_penetrant = 0.0
        total_ndt_cleaner = 0.0
        total_ndt_developer = 0.0
        total_ndt_fluorescent_pen = 0.0
        total_indiv_ot_amounts = [0] * 10
        has_note = False
        
        # Display aggregated entries
        for _, entry in grouped.iterrows():
            def clean_str(val):
                return str(val).replace('nan', '').replace('None', '').strip()
                
            mat_id = entry['MaterialID']
            
            # [ROBUST] Handle type mismatch in MaterialID lookup for Monthly Tab
            def get_mat_name(m_id):
                if pd.isna(m_id) or str(m_id).lower() == 'nan': return "N/A"
                # Try to find name from materials_df
                try:
                    # Convert both to float for numeric comparison
                    m_id_f = float(m_id)
                    matches = self.materials_df[pd.to_numeric(self.materials_df['MaterialID'], errors='coerce') == m_id_f]
                    if not matches.empty:
                        return str(matches.iloc[0]['품목명'])
                except: pass
                
                # Fallback to string comparison
                matches = self.materials_df[self.materials_df['MaterialID'].astype(str) == str(m_id)]
                if not matches.empty:
                    return str(matches.iloc[0]['품목명'])
                
                return f"ID: {m_id}"

            mat_name = get_mat_name(mat_id)
            
            # Apply material filter
            if filter_material != '전체' and mat_name != filter_material:
                continue
            
            # Get aggregated values from clean columns
            work_count = entry.get('WorkerCount', 0.0)
            ot_hours = entry.get('OT시간', 0.0)
            ot_amount = entry.get('OT금액', 0.0)
            test_amount = entry.get('검사량', 0.0)
            unit_price = entry.get('단가', 0.0)
            travel_cost = entry.get('출장비', 0.0)
            meal_cost = entry.get('일식', 0.0)
            test_fee = entry.get('검사비', 0.0)
            
            # RTK values
            rtk_center = entry.get('RTK_센터미스', 0.0)
            rtk_density = entry.get('RTK_농도', 0.0)
            rtk_marking = entry.get('RTK_마킹미스', 0.0)
            rtk_film = entry.get('RTK_필름마크', 0.0)
            rtk_handling = entry.get('RTK_취급부주의', 0.0)
            rtk_customer = entry.get('RTK_고객불만', 0.0)
            rtk_other = entry.get('RTK_기타', 0.0)
            rtk_total = rtk_center + rtk_density + rtk_marking + rtk_film + rtk_handling + rtk_customer + rtk_other
            
            # NDT values
            ndt_fluorescent_mag = entry.get('NDT_형광자분', 0.0)
            ndt_magnet = entry.get('NDT_자분', 0.0) + entry.get('NDT_흑색자분', 0.0)
            ndt_paint = entry.get('NDT_페인트', 0.0) + entry.get('NDT_백색페인트', 0.0)
            ndt_penetrant = entry.get('NDT_침투제', 0.0)
            ndt_cleaner = entry.get('NDT_세척제', 0.0)
            ndt_developer = entry.get('NDT_현상제', 0.0)
            ndt_fluorescent_pen = entry.get('NDT_형광', 0.0) + entry.get('NDT_형광침투제', 0.0)
            
            # Accumulate totals
            total_worker_count += work_count
            total_ot_hours += ot_hours
            total_ot_amount += ot_amount
            total_test_amount += test_amount
            total_unit_price += unit_price
            total_travel_cost += travel_cost
            total_meal_cost += meal_cost
            total_test_fee += test_fee
            
            total_rtk_center += rtk_center
            total_rtk_density += rtk_density
            total_rtk_marking += rtk_marking
            total_rtk_film += rtk_film
            total_rtk_handling += rtk_handling
            total_rtk_customer += rtk_customer
            total_rtk_other += rtk_other
            total_ndt_fluorescent_mag += ndt_fluorescent_mag
            total_ndt_magnet += ndt_magnet
            total_ndt_paint += ndt_paint
            total_ndt_penetrant += ndt_penetrant
            total_ndt_cleaner += ndt_cleaner
            total_ndt_developer += ndt_developer
            total_ndt_fluorescent_pen += ndt_fluorescent_pen
            
            # Extract worker names from User, User2, ..., User10
            all_workers = []
            # [ROBUST] Extract workers from matched columns
            for u_c, _, _ in matched_pairs:
                val = self.clean_nan(entry.get(u_c, ''))
                if val:
                    # Split in case it was already joined in aggregation
                    all_workers.extend([v.strip() for v in val.split(' | ') if v.strip()])
            worker_str = self.format_worker_summary(all_workers)
            
            # Concatenate work times
            all_worktimes = []
            worktime_cols = ['WorkTime', 'WorkTime2', 'WorkTime3', 'WorkTime4', 'WorkTime5', 'WorkTime6', 'WorkTime7', 'WorkTime8', 'WorkTime9', 'WorkTime10']
            for col in worktime_cols:
                val = str(entry.get(col, '')).strip()
                if val and val != 'nan' and val != '0.0':
                    all_worktimes.extend([v.strip() for v in val.split(' | ') if v.strip()])
            worktime_str = ", ".join(sorted(set(all_worktimes)))
            
            # Extract OT amounts only (suppress time strings)
            ot_values = []
            # [ROBUST] Pair-based OT extraction (and ensure exactly 10 items for column alignment)
            for i in range(1, 11):
                if i <= len(matched_pairs):
                    u_c, _, o_c = matched_pairs[i-1]
                    if not o_c:
                        ot_values.append('')
                        continue
                    val = str(entry.get(o_c, '')).strip()
                    
                    # [FIX] Eliminate ghost OT values in monthly view: Skip if no worker is assigned to this slot
                    user_val = str(entry.get(u_c, '')).strip()
                    has_worker = user_val and user_val != 'nan'
                    
                    if has_worker and val and val != 'nan' and val != '0.0':
                        # Handle multiple OTs if joined by aggregation separator ' | '
                        sub_vals = [v.strip() for v in val.split(' | ') if v.strip()]
                        parsed_ots = []
                        for v_str in sub_vals:
                            v_clean = v_str.replace(',', '')
                            if v_clean.isdigit() and int(v_clean) > 100:
                                # [PARITY] Monetary amount: Calculate hours from paired WorkTime
                                wt_val = str(entry.get(w_c, '')).strip() if w_c else ''
                                # Note: for individuals, we usually just show the amount in these columns
                                parsed_ots.append(f"{int(v_clean):,}")
                            elif '(' in v_str and '원)' in v_str:
                                try:
                                    amount_str = v_str.split('(')[1].split('원')[0].replace(',', '').strip()
                                    amount = int(float(amount_str))
                                    parsed_ots.append(f"{amount:,}")
                                except: pass
                            else:
                                try:
                                    amt = self.calculate_ot_amount(v_str)
                                    if amt > 0: parsed_ots.append(f"{amt:,}")
                                    elif any(x in v_str for x in [':', '시', '시간', '~', '-']): pass 
                                    else: parsed_ots.append(v_str)
                                except: pass
                        
                        ot_values.append(", ".join(parsed_ots))
                        try:
                            # [FIX] Sum ALL values accurately
                            for p_val in parsed_ots:
                                try:
                                    amt = int(p_val.replace(',', ''))
                                    total_indiv_ot_amounts[i-1] += amt
                                except: pass
                        except: pass
                    else:
                        ot_values.append('')
                else:
                    ot_values.append('')
            
            if clean_str(entry.get('Note', '')): has_note = True
            
            val_tuple = (
                int(entry['Year']),
                int(entry['Month']),
                entry.get('Site', ''),
                worker_str,
                worktime_str,
                f"{ot_hours:.1f}" if ot_hours > 0 else '',
                f"{ot_amount:,.0f}" if ot_amount > 0 else '',
                *ot_values,  # Index 7 to 16
                f"{test_amount:.1f}" if test_amount > 0 else '',
                f"{unit_price:,.0f}" if unit_price > 0 else '',
                f"{travel_cost:,.0f}" if travel_cost > 0 else '',
                f"{meal_cost:,.0f}" if meal_cost > 0 else '',
                f"{test_fee:,.0f}" if test_fee > 0 else '',
                mat_name, # Index 22: 품목명
                f"{rtk_center:.1f}" if rtk_center > 0 else '', # 23
                f"{rtk_density:.1f}" if rtk_density > 0 else '', # 24
                f"{rtk_marking:.1f}" if rtk_marking > 0 else '', # 25
                f"{rtk_film:.1f}" if rtk_film > 0 else '', # 26
                f"{rtk_handling:.1f}" if rtk_handling > 0 else '', # 27
                f"{rtk_customer:.1f}" if rtk_customer > 0 else '', # 28
                f"{rtk_other:.1f}" if rtk_other > 0 else '', # 29
                f"{rtk_total:.1f}" if rtk_total > 0 else '', # 30
                f"{ndt_fluorescent_mag:.1f}" if ndt_fluorescent_mag > 0 else '', # 31
                f"{ndt_magnet:.1f}" if ndt_magnet > 0 else '', # 32
                f"{ndt_paint:.1f}" if ndt_paint > 0 else '', # 33
                f"{ndt_penetrant:.1f}" if ndt_penetrant > 0 else '', # 34
                f"{ndt_cleaner:.1f}" if ndt_cleaner > 0 else '', # 35
                f"{ndt_developer:.1f}" if ndt_developer > 0 else '', # 36: 현상제
                f"{ndt_fluorescent_pen:.1f}" if ndt_fluorescent_pen > 0 else '', # 37
                '',  # Index 38: Note
                ", ".join(sorted(set(all_workers))) # Index 39: Full작업자
            )
            # [ROBUST] Final length check for 40 columns
            while len(val_tuple) < 40: val_tuple += ("",)
            
            self.monthly_usage_tree.insert('', tk.END, values=val_tuple)
            
            # [NEW] Popup sync
            if 'monthly' in self.detached_windows:
                p_tree = self.detached_windows['monthly']['tree']
                p_tree.insert('', tk.END, values=val_tuple) # Keep all 40 cols
        
        # Add total row at the bottom if there's data
        if not grouped.empty:
            total_rtk_sum = total_rtk_center + total_rtk_density + total_rtk_marking + total_rtk_film + total_rtk_handling + total_rtk_customer + total_rtk_other
            
            # Configure tag for total row
            self.monthly_usage_tree.tag_configure('total', background='#E8F4F8', font=('Arial', 12, 'bold'))
            
            total_values = (
                '',
                '',
                '--- 전체 누계 ---',
                '',  # 작업자
                f"{total_worker_count:.1f}" if total_worker_count > 0 else '', # 작업시간 (총 공수)
                f"{total_ot_hours:.1f}" if total_ot_hours > 0 else '',
                f"{total_ot_amount:,.0f}" if total_ot_amount > 0 else '',
                *[f"{a:,.0f}" if a > 0 else '' for a in total_indiv_ot_amounts],
                f"{total_test_amount:,.0f}" if total_test_amount > 0 else '',
                '', # 단가
                f"{total_travel_cost:,.0f}" if total_travel_cost > 0 else '',
                f"{total_meal_cost:,.0f}" if total_meal_cost > 0 else '',
                f"{total_test_fee:,.0f}" if total_test_fee > 0 else '',
                '', # 품목명
                f"{total_rtk_center:.1f}" if total_rtk_center > 0 else '',
                f"{total_rtk_density:.1f}" if total_rtk_density > 0 else '',
                f"{total_rtk_marking:.1f}" if total_rtk_marking > 0 else '',
                f"{total_rtk_film:.1f}" if total_rtk_film > 0 else '',
                f"{total_rtk_handling:.1f}" if total_rtk_handling > 0 else '',
                f"{total_rtk_customer:.1f}" if total_rtk_customer > 0 else '',
                f"{total_rtk_other:.1f}" if total_rtk_other > 0 else '',
                f"{total_rtk_sum:.1f}" if total_rtk_sum > 0 else '',
                f"{total_ndt_fluorescent_mag:.1f}" if total_ndt_fluorescent_mag > 0 else '',
                f"{total_ndt_magnet:.1f}" if total_ndt_magnet > 0 else '',
                f"{total_ndt_paint:.1f}" if total_ndt_paint > 0 else '',
                f"{total_ndt_penetrant:.1f}" if total_ndt_penetrant > 0 else '',
                f"{total_ndt_cleaner:.1f}" if total_ndt_cleaner > 0 else '',
                f"{total_ndt_developer:.1f}" if total_ndt_developer > 0 else '',
                f"{total_ndt_fluorescent_pen:.1f}" if total_ndt_fluorescent_pen > 0 else '',
                '', # 비고
                ''  # (Full작업자) hidden storage
            )
            self.monthly_usage_tree.insert('', tk.END, values=total_values, tags=('total',))
            
            # [NEW] Popup sync for total row
            if 'monthly' in self.detached_windows:
                p_tree = self.detached_windows['monthly']['tree']
                p_tree.insert('', tk.END, values=total_values, tags=('total',))
            
            # --- [NEW] Populate Summaries (Initial/Total) ---
            self._populate_monthly_summary_trees(df, has_note)
            
            # --- Dynamic Column Auto-Hide (same logic as Site tab) ---
            def is_active_m(val):
                if val is None: return False
                s = str(val).strip().lower()
                if s in ('', '0', '0.0', '0.00', 'nan', 'none', '-', '0.0시간', '0시간', '0원'): return False
                try:
                    import re as _rem
                    clean = _rem.sub(r'[^0-9\.\-]', '', s)
                    return bool(clean) and abs(float(clean)) > 0.001
                except:
                    return bool(s)

            all_cols = list(self.monthly_usage_tree['columns'])
            monthly_hide = set()
            
            # [REFINED] Minimal always_show list for smarter monthly column hiding
            always_show = {'연도', '월', '현장'}
            
            # Individual OT slots
            for i in range(1, 11):
                col = f'OT{i}' if i > 1 else 'OT시간'
                # Check from total amounts
                amt = total_indiv_ot_amounts[i-1] if i <= len(total_indiv_ot_amounts) else 0
                if not is_active_m(amt):
                    monthly_hide.add(f'OT{i}')
                    monthly_hide.add(f'작업자{i}' if i > 1 else '')
            
            # Cost columns
            if not is_active_m(total_travel_cost): monthly_hide.add('출장비')
            if not is_active_m(total_meal_cost): monthly_hide.add('일식')
            if not is_active_m(total_test_fee): monthly_hide.add('검사비')
            if not is_active_m(total_test_amount): monthly_hide.add('수량')
            if not is_active_m(total_ot_hours): monthly_hide.add('OT시간')
            if not is_active_m(total_ot_amount): monthly_hide.add('OT금액')
            
            # RTK columns
            rtk_totals = [total_rtk_center, total_rtk_density, total_rtk_marking, total_rtk_film, total_rtk_handling, total_rtk_customer, total_rtk_other]
            rtk_col_names_m = ['센터미스', '농도', '마킹미스', '필름마크', '취급부주의', '고객불만', '기타', 'RTK총계']
            for i, col in enumerate(rtk_col_names_m[:7]):
                if not is_active_m(rtk_totals[i]): monthly_hide.add(col)
            rtk_sum = sum(rtk_totals)
            if not is_active_m(rtk_sum): monthly_hide.add('RTK총계')
            
            # NDT columns
            ndt_totals = [total_ndt_fluorescent_mag, total_ndt_magnet, total_ndt_paint, total_ndt_penetrant, total_ndt_cleaner, total_ndt_developer, total_ndt_fluorescent_pen]
            ndt_col_names_m = ['형광자분', '흑색자분', '백색페인트', '침투제', '세척제', '현상제', '형광침투제']
            for i, col in enumerate(ndt_col_names_m):
                if not is_active_m(ndt_totals[i]): monthly_hide.add(col)
            
            # note column
            if not has_note: monthly_hide.add('비고')
            
            display_cols_m = [c for c in all_cols if c not in ('(Full작업자)',) and c not in monthly_hide]
            self.monthly_usage_tree['displaycolumns'] = display_cols_m
            if 'monthly' in self.detached_windows:
                self.detached_windows['monthly']['tree']['displaycolumns'] = display_cols_m
                
            # Ensure Total Row stays at bottom
            self.monthly_usage_tree.detach(self.monthly_usage_tree.get_children()[-1])
            self.monthly_usage_tree.insert('', tk.END, values=total_values, tags=('total',))
        else:
            # If empty data
            visible_cols = ['연도', '월', '현장', '작업자', '작업시간', '품목명']
            self.monthly_usage_tree['displaycolumns'] = visible_cols
            for col in visible_cols:
                self.monthly_usage_tree.column(col, stretch=False, minwidth=20)

    def on_monthly_usage_select(self, event):
        """Update site and worker summaries when a row is selected in monthly usage tree"""
        selection = self.monthly_usage_tree.selection()
        if not selection:
            return
            
        item = selection[0]
        values = self.monthly_usage_tree.item(item, 'values')
        tags = self.monthly_usage_tree.item(item, 'tags')
        
        # If no data stored yet or error
        if not hasattr(self, 'current_monthly_df') or self.current_monthly_df.empty:
            return
            
        # If total row is selected, show total summaries
        if 'total' in tags:
            self._populate_monthly_summary_trees(self.current_monthly_df)
            return
            
        # Extract row info (Year, Month, Site, Material)
        try:
            year = int(values[0])
            month = int(values[1])
            site = str(values[2]).strip()
            # Material is at index 22 (after 연도-월-현장-작업자-작업시간-OT시간-OT금액-OT1...OT10-검사량-단가-출장비-일식-검사비)
            mat_name = str(values[22]).strip()
            
            matching_ids = []
            # Find matching MateriaIDs for this mat_name from master data
            if hasattr(self, 'materials_df') and not self.materials_df.empty:
                matches = self.materials_df[self.materials_df['MaterialName'].astype(str).str.contains(mat_name, case=False, na=False)]
                if not matches.empty:
                    matching_ids = matches['MaterialID'].tolist()

            # Filter the current monthly dataset
            mask = (self.current_monthly_df['Year'] == year) & \
                   (self.current_monthly_df['Month'] == month) & \
                   (self.current_monthly_df['Site'] == site)

            
            # [ROBUST] Material Filter: matches 품목명 or direct MaterialID (for manual entries)
            if not matching_ids:
                # If not in master materials, check if mat_name itself exists as an ID in the data
                if mat_name in self.current_monthly_df['MaterialID'].astype(str).values:
                    matching_ids = [mat_name]
                else:
                    # Fallback: check case/space insensitive match in data
                    m_norm = mat_name.replace(' ', '').upper()
                    possible_ids = self.current_monthly_df['MaterialID'].dropna().unique()
                    for p_id in possible_ids:
                        if str(p_id).replace(' ', '').upper() == m_norm:
                            matching_ids.append(p_id)
            
            if matching_ids:
                mask = mask & (self.current_monthly_df['MaterialID'].isin(matching_ids))
            
            filtered_subset = self.current_monthly_df[mask]
            self._populate_monthly_summary_trees(filtered_subset)
            
        except Exception as e:
            print(f"DEBUG: Error in on_monthly_usage_select: {e}")

    def _populate_monthly_summary_trees(self, df, has_note=None):
        """Helper to fill site and worker summary trees with given data subset"""
        # Clear current views
        for item in self.site_summary_tree.get_children():
            self.site_summary_tree.delete(item)
        for item in self.worker_summary_tree.get_children():
            self.worker_summary_tree.delete(item)
            
        # [NEW] Clear detached views
        detached = self.detached_windows.get('monthly')
        if detached:
            for item in detached['site_tree'].get_children(): detached['site_tree'].delete(item)
            for item in detached['worker_tree'].get_children(): detached['worker_tree'].delete(item)
            
        if df.empty:
            return

        # --- Populate Site Summary ---
        # Include MaterialID in grouping and add relevant NDT/Cost fields to aggregation
        site_agg_dict = {
            '검사비': 'sum',
            '출장비': 'sum',
            'Usage': 'sum',
            '검사량': 'sum'
        }
        
        # Add RTK fields for RT매수 calculation
        rtk_fields = ['RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크', 'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타']
        for rf in rtk_fields:
            if rf in df.columns: 
                # [DEFENSIVE] Force numeric before aggregation
                df[rf] = pd.to_numeric(df[rf], errors='coerce').fillna(0.0)
                site_agg_dict[rf] = 'sum'
            
        # Add NDT fields
        ndt_map = {
            'NDT_형광자분': 'sum',
            'NDT_자분': 'sum',
            'NDT_흑색자분': 'sum',
            'NDT_페인트': 'sum',
            'NDT_백색페인트': 'sum',
            'NDT_침투제': 'sum',
            'NDT_세척제': 'sum',
            'NDT_현상제': 'sum',
            'NDT_형광': 'sum',
            'NDT_형광침투제': 'sum'
        }
        for nf in ndt_map:
            if nf in df.columns: 
                # [DEFENSIVE] Force numeric before aggregation
                df[nf] = pd.to_numeric(df[nf], errors='coerce').fillna(0.0)
                site_agg_dict[nf] = 'sum'

        # Helper for joining values in site-only summary
        def join_unique_non_empty(series):
            vals = [" ".join(str(v).split()) for v in series if pd.notna(v) and str(v).strip()]
            return " | ".join(sorted(set(vals)))
            
        # Add Joiner for mixed MaterialIDs in site-only summary
        site_agg_dict['MaterialID'] = join_unique_non_empty
        
        # Ensure Site and Method are strings and filled
        if 'Site' in df.columns: df['Site'] = df['Site'].fillna('').astype(str)
        if '검사방법' in df.columns: df['검사방법'] = df['검사방법'].fillna('미지정').astype(str)
        
        site_summary = df.groupby(['Site', '검사방법'], dropna=False).agg(site_agg_dict).reset_index()
        
        # [NEW] Track active columns for dynamic hiding
        site_cols = ('현장', '검사방법', '품목명', '수량', '검사비', '출장비', '형광자분', '흑색자분', '백색페인트', 
                     '침투제', '세척제', '현상제', '형광침투제', '센터미스', '농도', '마킹미스', '필름마크', '취급부주의', '고객불만', '기타', 'RTK총계')
        active_cols = set(['현장', '검사방법', '품목명', '수량']) # Mandatory columns
        
        # Define a robust active check for numeric data
        def is_active(val):
            if val is None: return False
            s = str(val).strip().lower()
            if s in ('', '0', '0.0', '0.00', 'nan', 'none', '-'):
                return False
            try:
                # Remove common separators and units for numeric check
                v = float(s.replace(',', '').replace('원', '').replace('시간', '').strip())
                return abs(v) > 0.001
            except:
                return bool(s)

        for _, row in site_summary.iterrows():
            # Get RT film usage from unified Usage column
            film_total = row.get('Usage', 0.0)
            
            # Calculate values
            black_mag = row.get('NDT_자분', 0.0) + row.get('NDT_흑색자분', 0.0)
            white_paint = row.get('NDT_페인트', 0.0) + row.get('NDT_백색페인트', 0.0)
            pt_pen = row.get('NDT_침투제', 0.0)
            pt_cln = row.get('NDT_세척제', 0.0)
            pt_dev = row.get('NDT_현상제', 0.0)
            pt_fluoro = row.get('NDT_형광', 0.0) + row.get('NDT_형광침투제', 0.0)
            
            r_center = row.get('RTK_센터미스', 0.0)
            r_density = row.get('RTK_농도', 0.0)
            r_marking = row.get('RTK_마킹미스', 0.0)
            r_film = row.get('RTK_필름마크', 0.0)
            r_careless = row.get('RTK_취급부주의', 0.0)
            r_customer = row.get('RTK_고객불만', 0.0)
            r_other = row.get('RTK_기타', 0.0)
            rtk_sum = r_center + r_density + r_marking + r_film + r_careless + r_customer + r_other
            
            # Track which columns are active using robust threshold
            if is_active(row['검사비']): active_cols.add('검사비')
            if is_active(row['출장비']): active_cols.add('출장비')
            if is_active(row.get('NDT_형광자분', 0.0)): active_cols.add('형광자분')
            if is_active(black_mag): active_cols.add('흑색자분')
            if is_active(white_paint): active_cols.add('백색페인트')
            if is_active(pt_pen): active_cols.add('침투제')
            if is_active(pt_cln): active_cols.add('세척제')
            if is_active(pt_dev): active_cols.add('현상제')
            if is_active(pt_fluoro): active_cols.add('형광침투제')
            if is_active(r_center): active_cols.add('센터미스')
            if is_active(r_density): active_cols.add('농도')
            if is_active(r_marking): active_cols.add('마킹미스')
            if is_active(r_film): active_cols.add('필름마크')
            if is_active(r_careless): active_cols.add('취급부주의')
            if is_active(r_customer): active_cols.add('고객불만')
            if is_active(r_other): active_cols.add('기타')
            if is_active(rtk_sum): active_cols.add('RTK총계')
            
            # Get Material Name (Handle multiple joined IDs if site grouping consolidated them)
            mat_ids_raw = str(row['MaterialID']).split(' | ')
            mat_names_list = []
            for m_id_str in mat_ids_raw:
                curr_id = m_id_str.strip()
                if not curr_id or curr_id.lower() == 'nan': continue
                
                # [ROBUST] Logic to find material name from ID (handling numeric conversion)
                def get_single_mat_name(m_id_val):
                    try: 
                        m_f = float(m_id_val)
                        matches = self.materials_df[pd.to_numeric(self.materials_df['MaterialID'], errors='coerce') == m_f]
                        if not matches.empty: return str(matches.iloc[0]['품목명'])
                    except: pass
                    matches = self.materials_df[self.materials_df['MaterialID'].astype(str) == str(m_id_val)]
                    if not matches.empty: return str(matches.iloc[0]['품목명'])
                    return f"ID:{m_id_val}"
                
                mat_names_list.append(get_single_mat_name(curr_id))
                    
            mat_name = ", ".join(sorted(set(mat_names_list))) if mat_names_list else "품목 미지정"

            values = (
                row['Site'],
                row['검사방법'],
                mat_name,
                f"{row.get('검사량', 0.0):.1f}" if is_active(row.get('검사량', 0.0)) else '',
                f"{row['검사비']:,.0f}" if is_active(row['검사비']) else '',
                f"{row['출장비']:,.0f}" if is_active(row['출장비']) else '',
                f"{row.get('NDT_형광자분', 0.0):.1f}" if is_active(row.get('NDT_형광자분', 0.0)) else '',
                f"{black_mag:.1f}" if is_active(black_mag) else '',
                f"{white_paint:.1f}" if is_active(white_paint) else '',
                f"{pt_pen:.1f}" if is_active(pt_pen) else '',
                f"{pt_cln:.1f}" if is_active(pt_cln) else '',
                f"{pt_dev:.1f}" if is_active(pt_dev) else '',
                f"{pt_fluoro:.1f}" if is_active(pt_fluoro) else '',
                f"{r_center:.1f}" if is_active(r_center) else '',
                f"{r_density:.1f}" if is_active(r_density) else '',
                f"{r_marking:.1f}" if is_active(r_marking) else '',
                f"{r_film:.1f}" if is_active(r_film) else '',
                f"{r_careless:.1f}" if is_active(r_careless) else '',
                f"{r_customer:.1f}" if is_active(r_customer) else '',
                f"{r_other:.1f}" if is_active(r_other) else '',
                f"{rtk_sum:.1f}" if is_active(rtk_sum) else ''
            )
            self.site_summary_tree.insert('', tk.END, values=values)
            if detached:
                detached['site_tree'].insert('', tk.END, values=values)

        # Apply dynamic column hiding to Site Summary
        visible_sum_cols = [col for col in site_cols if col in active_cols]
        self.site_summary_tree['displaycolumns'] = visible_sum_cols
        if detached:
            detached['site_tree']['displaycolumns'] = visible_sum_cols

        # --- Populate Worker Summary ---
        worker_data = []
        
        # [ROBUST] Use the same precise column detection as update_monthly_usage_view
        def find_paired_cols_local(cols):
            pairs = []
            col_set = set(cols)
            for i in range(1, 11):
                u_n = 'User' if i == 1 else f'User{i}'
                w_n = 'WorkTime' if i == 1 else f'WorkTime{i}'
                o_n = 'OT' if i == 1 else f'OT{i}'
                if u_n in col_set:
                    pairs.append((u_n, w_n if w_n in col_set else None, o_n if o_n in col_set else None))
            return pairs

        df_norm = df.copy()
        df_norm.columns = [str(c).strip().replace(' ', '') for c in df_norm.columns]
        
        pairs = find_paired_cols_local(df_norm.columns)
        for uc, wc, oc in pairs:
            temp_df = pd.DataFrame()
            temp_df['WorkerName'] = df_norm[uc]
            temp_df['ShiftType'] = df_norm[wc] if wc else '주간'
            temp_df['OTValue'] = df_norm[oc] if oc else ''
            
            # Context for deduping
            temp_df['Date'] = df_norm['Date'] if 'Date' in df_norm.columns else None
            temp_df['Site'] = df_norm['Site'] if 'Site' in df_norm.columns else ''
            temp_df['MaterialID'] = df_norm['MaterialID'] if 'MaterialID' in df_norm.columns else ''
            temp_df['WorkTime'] = df_norm['WorkTime'] if 'WorkTime' in df_norm.columns else '' # Base worktime for deduping
            
            worker_data.append(temp_df)
        
        if worker_data:
            # [NEW] Track active columns for dynamic hiding in Worker Summary
            worker_sum_cols = ('작업자', '총공수', '연장(시간)', '야간(시간)', '휴일(시간)', '총OT(시간)', '연장(금액)', '야간(금액)', '휴일(금액)', '총OT(금액)')
            active_worker_cols = set(['작업자', '총공수'])
            
            worker_df = pd.concat(worker_data)
            worker_df['WorkerName'] = worker_df['WorkerName'].apply(self.clean_nan)
            worker_df = worker_df[worker_df['WorkerName'] != '']
            
            # [NEW] Worker-level Deduping to prevent double-counting in split records
            if not worker_df.empty:
                def make_worker_key(row):
                    d = self._safe_format_datetime(row.get('Date', ''), '%Y-%m-%d')
                    s = str(row.get('Site', '')).strip()
                    w = str(row.get('WorkTime', '')).strip()
                    m = str(row.get('MaterialID', '')).strip()
                    un = str(row.get('WorkerName', '')).strip()
                    return (d, s, w, m, un)
                
                worker_df['_dedupe_key'] = worker_df.apply(make_worker_key, axis=1)
                # Keep the first instance of a worker in any given activity unit
                worker_df = worker_df.drop_duplicates(subset=['_dedupe_key']).drop(columns=['_dedupe_key'])
            
            if not worker_df.empty:
                # [FIXED] Use row['Date'] if available, otherwise fallback to first available date in subset
                def get_date_val(row):
                    if 'Date' in row and pd.notna(row['Date']): return row['Date']
                    try: return df.iloc[0]['Date']
                    except: return None

                # [NEW] Enhanced OT hour extraction: check WorkTime if OTValue is only amount
                def get_ot_hours_robust(row):
                    ot_val = str(row['OTValue']).strip()
                    if not ot_val or ot_val == '0': return 0.0
                    
                    # If it's already "N시간...", parse it
                    if '시간' in ot_val:
                        return self._parse_ot_hours(ot_val)
                    
                    # If it's just an amount, calculate hours from WorkTime
                    if ot_val.replace(',', '').isdigit() and int(ot_val.replace(',', '')) > 100:
                        wt_val = row['ShiftType']
                        date_val = get_date_val(row)
                        h, _ = self._calculate_ot_from_worktime(wt_val, date_val)
                        return h
                    
                    return self._parse_ot_hours(ot_val)

                worker_df['OT_H'] = worker_df.apply(get_ot_hours_robust, axis=1)
                
                # Split calculation also needs to be robust
                def get_split_robust(row):
                    ot_val = str(row['OTValue']).strip()
                    date_val = get_date_val(row)
                    if '시간' in ot_val or not (ot_val.replace(',', '').isdigit() and int(ot_val.replace(',', '')) > 100):
                        return self._calculate_split_ot_hours(ot_val, date_val)
                    else:
                        # Recalculate from WorkTime
                        wt_val = row['ShiftType']
                        
                        # We need the simulation logic from _calculate_split_ot_hours but using wt_val
                        # Let's use a temporary string that _calculate_split_ot_hours can parse
                        h, _ = self._calculate_ot_from_worktime(wt_val, date_val)
                        fake_ot_str = f"{h}시간 ({ot_val}원)"
                        # Check start hour from wt_val for range logic
                        marker_pattern = MARKER_PATTERN
                        clean_wt = marker_pattern.sub('', str(wt_val)).strip()
                        if '~' in clean_wt:
                            start_time = clean_wt.split('~')[0]
                            fake_ot_str = f"{start_time}~{h}시간 ({ot_val}원)"
                        
                        return self._calculate_split_ot_hours(fake_ot_str, date_val)

                # Now returns 3 values (day, night, holiday_dawn)
                worker_df[['OT_H_Day', 'OT_H_Night', 'OT_H_Holiday_Dawn']] = worker_df.apply(
                    lambda r: pd.Series(get_split_robust(r)), axis=1
                )
                worker_df['OT_A'] = worker_df['OTValue'].apply(self.calculate_ot_amount)
                
                # [FIXED] Parse fractional work counts from ShiftType (e.g., "0.5주간" -> 0.5)
                def parse_work_count(val):
                    s = str(val).strip()
                    if not s: return 1.0
                    try:
                        # Extract first float/int found in string
                        match = _re.search(r'([0-9\.]+)', s)
                        if match: return float(match.group(1))
                    except: pass
                    return 1.0
                
                worker_df['Count'] = worker_df['ShiftType'].apply(parse_work_count)
                
                # [FIXED] Shift types: Automatically treat Saturday/Sunday as Holiday shifts
                def determine_shift_holiday(row):
                    s_type = str(row['ShiftType']).strip()
                    if '휴일' in s_type: return 1.0
                    d_val = get_date_val(row)
                    if d_val:
                        try:
                            if pd.to_datetime(d_val).weekday() >= 5: return 1.0
                        except: pass
                    return 0.0

                worker_df['Shift_Holiday'] = worker_df.apply(determine_shift_holiday, axis=1)
                worker_df['Shift_Day'] = worker_df['ShiftType'].apply(lambda x: 1.0 if '주간' in str(x) and '휴일' not in str(x) else 0.0)
                worker_df['Shift_Night'] = worker_df['ShiftType'].apply(lambda x: 1.0 if '야간' in str(x) and '휴일' not in str(x) else 0.0)
                
                # If explicit holiday shift, all OT is holiday.
                # If NOT explicit holiday shift, add the OT_H_Holiday_Dawn (Friday dawn) to holiday hours.
                worker_df['H_Holiday'] = worker_df.apply(lambda r: r['OT_H'] if r['Shift_Holiday'] > 0 else r['OT_H_Holiday_Dawn'], axis=1)
                
                # Calculate amount properly. If it was fully holiday, it's just OT_A.
                # If it was Friday Dawn, add the calculated dawn hours * 7500 to whatever other OT they had?
                # Actually, A_Holiday should just be the amount from those hours
                worker_df['A_Holiday'] = worker_df.apply(lambda r: r['OT_A'] if r['Shift_Holiday'] > 0 else r['OT_H_Holiday_Dawn'] * 7500, axis=1)
                
                # For day/night, they only get values if it's NOT a full holiday shift.
                worker_df['H_Day'] = worker_df.apply(lambda r: r['OT_H_Day'] if r['Shift_Holiday'] == 0 else 0.0, axis=1)
                worker_df['H_Night'] = worker_df.apply(lambda r: r['OT_H_Night'] if r['Shift_Holiday'] == 0 else 0.0, axis=1)
                worker_df['A_Day'] = worker_df.apply(lambda r: r['H_Day'] * 4000 if r['Shift_Holiday'] == 0 else 0.0, axis=1)
                worker_df['A_Night'] = worker_df.apply(lambda r: r['H_Night'] * 5000 if r['Shift_Holiday'] == 0 else 0.0, axis=1)

                # [REVERTED] Back to entry-based count
                worker_summary = worker_df.groupby('WorkerName').agg({
                    'Count': 'sum',
                    'H_Day': 'sum',
                    'H_Night': 'sum',
                    'H_Holiday': 'sum',
                    'OT_H': 'sum',
                    'A_Day': 'sum',
                    'A_Night': 'sum',
                    'A_Holiday': 'sum',
                    'OT_A': 'sum'
                }).reset_index()
                
                for _, row in worker_summary.iterrows():
                    # Track active columns using robust is_active helper
                    if is_active(row['H_Day']): active_worker_cols.add('연장(시간)')
                    if is_active(row['H_Night']): active_worker_cols.add('야간(시간)')
                    if is_active(row['H_Holiday']): active_worker_cols.add('휴일(시간)')
                    if is_active(row['OT_H']): active_worker_cols.add('총OT(시간)')
                    if is_active(row['A_Day']): active_worker_cols.add('연장(금액)')
                    if is_active(row['A_Night']): active_worker_cols.add('야간(금액)')
                    if is_active(row['A_Holiday']): active_worker_cols.add('휴일(금액)')
                    if is_active(row['OT_A']): active_worker_cols.add('총OT(금액)')
 
                    values = (
                        row['WorkerName'],
                        f"{row['Count']:.1f}" if is_active(row['Count']) else "",
                        f"{row['H_Day']:.1f}" if is_active(row['H_Day']) else "",
                        f"{row['H_Night']:.1f}" if is_active(row['H_Night']) else "",
                        f"{row['H_Holiday']:.1f}" if is_active(row['H_Holiday']) else "",
                        f"{row['OT_H']:.1f}" if is_active(row['OT_H']) else "",
                        f"{row['A_Day']:,.0f}" if is_active(row['A_Day']) else "",
                        f"{row['A_Night']:,.0f}" if is_active(row['A_Night']) else "",
                        f"{row['A_Holiday']:,.0f}" if is_active(row['A_Holiday']) else "",
                        f"{row['OT_A']:,.0f}" if is_active(row['OT_A']) else ""
                    )
                    self.worker_summary_tree.insert('', tk.END, values=values)
                    if detached:
                        detached['worker_tree'].insert('', tk.END, values=values)
                
                # Apply dynamic column hiding to Worker Summary
                visible_worker_cols = [col for col in worker_sum_cols if col in active_worker_cols]
                self.worker_summary_tree['displaycolumns'] = visible_worker_cols
                if detached:
                    detached['worker_tree']['displaycolumns'] = visible_worker_cols

    def apply_worker_shift_hours_to_budget(self):
        """월별 탭의 작업자별 누계(주간/야간/휴일) 시간을 공사실행예산서 특별근무 투입시간에 적용한다."""
        if not hasattr(self, 'labor_detail_widget'):
            messagebox.showwarning("공사탭 미초기화",
                                   "공사실행예산서 탭을 먼저 열어 초기화하세요.")
            return

        # worker_summary_tree에서 모든 행의 주간/야간/휴일 합산
        total_h_day = 0.0
        total_h_night = 0.0
        total_h_holiday = 0.0
        worker_count_day = 0
        worker_count_night = 0
        worker_count_holiday = 0

        def _f(val):
            try: return float(str(val).replace(',', '') or 0)
            except: return 0.0

        for item in self.worker_summary_tree.get_children():
            vals = self.worker_summary_tree.item(item, 'values')
            # 컬럼 순서: 작업자, 총공수, 연장(시간), 야간(시간), 휴일(시간), 총OT(시간), 연장(금액), 야간(금액), 휴일(금액), 총OT(금액)
            h_day     = _f(vals[2]) if len(vals) > 2 else 0.0
            h_night   = _f(vals[3]) if len(vals) > 3 else 0.0
            h_holiday = _f(vals[4]) if len(vals) > 4 else 0.0
            if h_day > 0:
                total_h_day += h_day
                worker_count_day += 1
            if h_night > 0:
                total_h_night += h_night
                worker_count_night += 1
            if h_holiday > 0:
                total_h_holiday += h_holiday
                worker_count_holiday += 1

        if total_h_day == 0 and total_h_night == 0 and total_h_holiday == 0:
            messagebox.showinfo("데이터 없음",
                                "작업자별 누계에 주간/야간/휴일 시간 데이터가 없습니다.\n"
                                "월별 탭에서 조회를 먼저 실행하세요.")
            return

        # LaborCostDetailWidget 특별근무 섹션에 입력
        ldw = self.labor_detail_widget
        # 연장근무 ← 주간 OT 시간 합계 / 인원수
        if worker_count_day > 0:
            avg_day = total_h_day / worker_count_day
            ldw.entries["연장근무"]['personnel'].delete(0, 'end')
            ldw.entries["연장근무"]['personnel'].insert(0, f"{worker_count_day:g}")
            ldw.entries["연장근무"]['period'].delete(0, 'end')
            ldw.entries["연장근무"]['period'].insert(0, f"{avg_day:.1f}")
        # 야간근무 ← 야간 OT 시간 합계
        if worker_count_night > 0:
            avg_night = total_h_night / worker_count_night
            ldw.entries["야간근무"]['personnel'].delete(0, 'end')
            ldw.entries["야간근무"]['personnel'].insert(0, f"{worker_count_night:g}")
            ldw.entries["야간근무"]['period'].delete(0, 'end')
            ldw.entries["야간근무"]['period'].insert(0, f"{avg_night:.1f}")
        # 휴일근무 ← 휴일 OT 시간 합계
        if worker_count_holiday > 0:
            avg_holiday = total_h_holiday / worker_count_holiday
            ldw.entries["휴일근무"]['personnel'].delete(0, 'end')
            ldw.entries["휴일근무"]['personnel'].insert(0, f"{worker_count_holiday:g}")
            ldw.entries["휴일근무"]['period'].delete(0, 'end')
            ldw.entries["휴일근무"]['period'].insert(0, f"{avg_holiday:.1f}")

        # 계산 반영
        ldw.calculate_all()

        # 공사실행예산서 탭으로 이동
        try:
            tab_idx = [self.notebook.tab(i, 'text') for i in range(self.notebook.index('end'))].index('공사실행예산서')
            self.notebook.select(tab_idx)
        except:
            pass

        messagebox.showinfo("적용 완료",
                            f"특별근무 투입시간 적용 완료!\n"
                            f"  연장근무: {worker_count_day}명 / {total_h_day:.1f}h (평균 {total_h_day/max(worker_count_day,1):.1f}h/인)\n"
                            f"  야간근무: {worker_count_night}명 / {total_h_night:.1f}h (평균 {total_h_night/max(worker_count_night,1):.1f}h/인)\n"
                            f"  휴일근무: {worker_count_holiday}명 / {total_h_holiday:.1f}h (평균 {total_h_holiday/max(worker_count_holiday,1):.1f}h/인)")

    def export_monthly_usage_history(self):
        """Export monthly usage data, site summaries, and worker summaries to a multi-sheet Excel file"""
        try:
            # 1. Collect Main Monthly Usage Data
            columns = self.monthly_usage_tree['columns']
            monthly_data = []
            for item in self.monthly_usage_tree.get_children():
                monthly_data.append(self.monthly_usage_tree.item(item, 'values'))
            
            if not monthly_data:
                messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
                return

            # 2. Collect Site Summary Data
            site_columns = self.site_summary_tree['columns']
            site_data = []
            for item in self.site_summary_tree.get_children():
                site_data.append(self.site_summary_tree.item(item, 'values'))

            # 3. Collect Worker Summary Data
            worker_columns = self.worker_summary_tree['columns']
            worker_data = []
            for item in self.worker_summary_tree.get_children():
                worker_data.append(self.worker_summary_tree.item(item, 'values'))
            
            # Prepare filename
            today = datetime.datetime.now().strftime('%Y%m%d')
            # Using a descriptive filename indicating it's a comprehensive report (종합)
            filename = f"월별집계_종합_{today}.xlsx"
            
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=filename,
                title="월별 집계(종합) 내역 저장",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not save_path:
                return

            # --- Process Monthly Usage Tree (Sheet 1) ---
            columns_to_export = [c for c in columns if c != '(Full작업자)']
            col_to_idx = {col: i for i, col in enumerate(columns)}
            
            final_monthly_data = []
            for row in monthly_data:
                new_row = list(row)
                # Swap summarized worker with full list if (Full작업자) exists
                if '작업자' in col_to_idx and '(Full작업자)' in col_to_idx:
                    full_idx = col_to_idx['(Full작업자)']
                    if full_idx < len(new_row):
                        full_val = new_row[full_idx]
                        if full_val:
                            new_row[col_to_idx['작업자']] = full_val
                
                # Filter only display columns
                final_row = [new_row[i] for i, c in enumerate(columns) if c != '(Full작업자)' and i < len(new_row)]
                final_monthly_data.append(final_row)
            
            monthly_df = pd.DataFrame(final_monthly_data, columns=columns_to_export)
            monthly_df = self.clean_df_export(monthly_df)

            # --- Process Site Summary (Sheet 2) ---
            site_df = pd.DataFrame(site_data, columns=site_columns)
            site_df = self.clean_df_export(site_df)

            # --- Process Worker Summary (Sheet 3) ---
            worker_df = pd.DataFrame(worker_data, columns=worker_columns)
            worker_df = self.clean_df_export(worker_df)

            # --- Save to Excel with Multiple Sheets and AutoFit ---
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                # Local helper for autofitting to avoid polluting global namespace
                def write_sheet(df, sheet_name):
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    worksheet = writer.sheets[sheet_name]
                    
                    def get_display_width(s):
                        width = 0
                        for char in str(s):
                            if ord(char) > 127: width += 2 # Double width for Korean
                            else: width += 1
                        return width

                    for idx, col in enumerate(df.columns):
                        series = df[col].astype(str)
                        lengths = series.apply(get_display_width)
                        max_val_len = lengths.max() if not lengths.empty else 0
                        header_len = get_display_width(col)
                        # Cap at 40 to allow better horizontal density in summaries
                        final_width = min(max(max_val_len, header_len) + 2, 40)
                        
                        col_letter = worksheet.cell(row=1, column=idx+1).column_letter
                        worksheet.column_dimensions[col_letter].width = final_width

                write_sheet(monthly_df, "월별집계_내역")
                write_sheet(site_df, "현장별_누계")
                write_sheet(worker_df, "작업자별_누계")

            messagebox.showinfo("완료", "월별 집계 내역(종합)이 저장되었습니다.\n(내역, 현장별, 작업자별 3개 시트 포함)")
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패: {e}")
            traceback.print_exc()

    def create_draggable_container(self, parent, label_text, widget_class, config_key, manage_list_key=None, grid_info=None, **widget_kwargs):
        """Create a draggable container with a label and a widget, styled as a visible box"""
        box_border_color = (
            getattr(self, 'style', None).lookup('TLabelframe', 'bordercolor') if hasattr(self, 'style') else ''
        ) or (
            getattr(self, 'style', None).lookup('TFrame', 'bordercolor') if hasattr(self, 'style') else ''
        ) or '#c0c0c0'

        container = tk.Frame(
            parent,
            relief="solid",
            borderwidth=1,
            highlightthickness=1,
            highlightbackground=box_border_color,
            bg=self.theme_bg
        )
        
        # Header container for label and buttons
        hdr = ttk.Frame(container) 
        hdr.pack(side='top', fill='x', padx=0, pady=0)
        
        # Drag handle icon
        lbl_drag = ttk.Label(hdr, text="✥", font=('Arial', 9), cursor='fleur')
        lbl_drag.pack(side='left', padx=1)
        self.make_header_draggable(lbl_drag, container)
        
        # Label
        lbl = ttk.Label(hdr, text=label_text, font=('Malgun Gothic', 9, 'bold'))
        lbl.pack(side='left', padx=1)
        self.make_header_draggable(lbl, container)
        
        # Icons container on the right
        btn_box = ttk.Frame(hdr)
        btn_box.pack(side='right', padx=1)

        # Rename icon
        btn_rename = ttk.Label(btn_box, text="✏️", font=('Arial', 8), cursor='hand2')
        btn_rename.pack(side='left', padx=1)
        btn_rename.bind('<Button-1>', lambda e: self.rename_widget_label(config_key))
        
        # Clone icon
        btn_clone = ttk.Label(btn_box, text="📋", font=('Arial', 8), cursor='hand2')
        btn_clone.pack(side='left', padx=1)
        btn_clone.bind('<Button-1>', lambda e: self.clone_widget(config_key))

        # Delete icon (X)
        btn_del = ttk.Label(btn_box, text="❌", font=('Arial', 8), cursor='hand2')
        btn_del.pack(side='left', padx=1)
        btn_del.bind('<Button-1>', lambda e: self.remove_box(config_key))

        # Internal Content Area
        content_area = ttk.Frame(container, padding=(1, 0))
        content_area.pack(side='top', fill='both', expand=True)

        # Widget
        widget = widget_class(content_area, **widget_kwargs)
        widget.pack(side='left', fill='both', expand=True)

        # Manage List Icon (Gear) - if it's a list-based widget
        if manage_list_key:
            btn_manage = ttk.Label(btn_box, text="⚙️", font=('Arial', 8), cursor='hand2')
            btn_manage.pack(side='left', padx=1)
            # [FIX] Bind AFTER widget is defined
            btn_manage.bind('<Button-1>', lambda e, cb=widget: self.open_list_management_dialog(manage_list_key, target_cb=cb))
        
        # If the widget is a basic tk widget (Text, Canvas), set its background
        if hasattr(widget, 'config') and 'bg' in widget.keys():
            try: widget.config(bg=self.theme_bg)
            except: pass
        
        # NEW: Handle grid placement first if grid_info provided
        if grid_info:
            container.grid(**grid_info)
        
        # Track for layout reset/config
        container._config_key = config_key
        container._label_widget = lbl
        container._widget = widget
        container._widget_class = widget_class
        container._widget_kwargs = widget_kwargs
        container._manage_list_key = manage_list_key
        
        # Register and make draggable (this now captures the CORRECT grid info)
        self.draggable_items[config_key] = container
        self.make_draggable(container, config_key)
        
        return container, widget

    def open_list_management_dialog(self, title_or_key, data_list=None, config_key=None, target_cb=None):
        """Open a generic dialog to manage (edit/delete) items in a data list"""
        if self.layout_locked: return

        # If data_list is None, we assume title_or_key is the config_key
        if data_list is None:
            config_key = title_or_key
            data_map = {
                'sites': ('현장 목록 관리', self.sites),
                'users': ('담당자 목록 관리', getattr(self, 'users', [])),
                'equipments': ('장비 목록 관리', getattr(self, 'equipments', [])),
                'vehicles': ('차량 목록 관리', getattr(self, 'vehicles', [])),
                'companies': ('업체 목록 관리', getattr(self, 'companies', [])),
                'materials': ('품목 목록 관리 (기본)', getattr(self, 'carestream_films', [])),
                'daily_units': ('단위 목록 관리', self.daily_units),
                'test_items': ('검사품명 목록 관리', getattr(self, 'test_items', [])),
                'applied_codes': ('적용코드 목록 관리', getattr(self, 'applied_codes', []))
            }
            if config_key not in data_map: return
            title, data_list = data_map[config_key]
        else:
            # Traditional 3-argument call: title, data_list, config_key
            title = title_or_key
            if config_key is None: config_key = title # Fallback

        # ── 이미 열린 창이 있으면 앞으로 가져오고 종료 ──────────────
        if not hasattr(self, '_list_mgmt_dialogs'):
            self._list_mgmt_dialogs = {}
        existing = self._list_mgmt_dialogs.get(config_key)
        if existing and existing.winfo_exists():
            existing.lift()
            existing.focus_set()
            return

        dialog = tk.Toplevel(self.root)
        self._list_mgmt_dialogs[config_key] = dialog  # 창 추적 등록
        dialog.title(title)
        dialog.geometry("600x450")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding=10)
        frame.pack(fill='both', expand=True)
        
        # [NEW] Search box for the list
        search_f = ttk.Frame(frame)
        search_f.pack(fill='x', pady=(0, 5))
        ttk.Label(search_f, text="🔍:").pack(side='left', padx=2)
        search_var = tk.StringVar()
        search_ent = ttk.Entry(search_f, textvariable=search_var)
        search_ent.pack(side='left', fill='x', expand=True, padx=2)
        search_ent.focus_set()

        # Add scrollbar
        scrollbar = ttk.Scrollbar(frame)
        scrollbar.pack(side='right', fill='y')
        
        listbox = tk.Listbox(frame, font=('Arial', 10))
        listbox.pack(fill='both', expand=True, side='left')
        listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=listbox.yview)
        
        def refresh_list(*args):
            query = search_var.get().strip().lower()
            listbox.delete(0, 'end')
            # Use original data_list for filtering
            for item in sorted(data_list):
                if not query or query in item.lower():
                    listbox.insert('end', item)
            # Trigger app-wide update of related comboboxes
            self.refresh_ui_for_list_change(config_key)

        search_var.trace_add("write", refresh_list)
        refresh_list()
            
        def select_and_close():
            sel = listbox.curselection()
            if not sel: return
            val = listbox.get(sel[0])
            if target_cb:
                # [FIX] Support both Entry and Combobox widgets
                if hasattr(target_cb, 'set'):
                    target_cb.set(val)
                else:
                    target_cb.delete(0, tk.END)
                    target_cb.insert(0, val)
                # Manually trigger any search/filter logic if needed
                if hasattr(self, 'update_daily_usage_view'):
                    self.update_daily_usage_view()
            dialog.destroy()

        listbox.bind("<Double-1>", lambda e: select_and_close())
            
        btn_frame = ttk.Frame(dialog, padding=5)
        btn_frame.pack(fill='x')
        
        # [NEW] Select button (Shows only if target_cb is present, or always as an alternative to double-click)
        ttk.Button(btn_frame, text="✅ 선택", command=select_and_close).pack(side='left', padx=2)

        def add_item():
            new_val = simpledialog.askstring("추가", "새 이름을 입력하세요:")
            if new_val and new_val.strip():
                val = new_val.strip()
                if val not in data_list:
                    data_list.append(val)
                    data_list.sort()
                    self.save_tab_config()
                    refresh_list()
                else:
                    messagebox.showinfo("정보", "이미 목록에 있는 이름입니다.")

        def edit_item():
            sel = listbox.curselection()
            if not sel: return
            idx = sel[0]
            old_val = listbox.get(idx)
            new_val = simpledialog.askstring("수정", f"[{old_val}] -> 새 이름을 입력하세요:", initialvalue=old_val)
            if new_val and new_val.strip() and new_val != old_val:
                if old_val in data_list: data_list.remove(old_val)
                data_list.append(new_val.strip())
                data_list.sort()
                self.save_tab_config()
                refresh_list()

        def delete_item():
            sel = listbox.curselection()
            if not sel: return
            idx = sel[0]
            val = listbox.get(idx)
            if messagebox.askyesno("삭제 확인", f"'{val}'을 목록에서 삭제하시겠습니까?"):
                if val in data_list: data_list.remove(val)
                # [NEW] If the deleted value is currently selected in the target combobox, clear it
                if target_cb:
                    current_val = target_cb.get().strip()
                    if current_val == val.strip():
                        if hasattr(target_cb, 'set'):
                            target_cb.set('')
                        else:
                            target_cb.delete(0, tk.END)
                self.save_tab_config()
                refresh_list()

        def on_close():
            self._list_mgmt_dialogs.pop(config_key, None)
            dialog.destroy()

        dialog.protocol("WM_DELETE_WINDOW", on_close)

        def bulk_register_to_stock():
            if messagebox.askyesno("일괄 등록 확인", "현재 목록에 있는 모든 품목을 '재고 현황' 데이터베이스에 새로 등록하시겠습니까?\n(이미 등록된 품목은 중복 등록되지 않습니다.)"):
                self.bulk_register_materials_to_stock(data_list)

        ttk.Button(btn_frame, text="추가", command=add_item).pack(side='left', padx=5, expand=True)
        ttk.Button(btn_frame, text="수정", command=edit_item).pack(side='left', padx=5, expand=True)
        ttk.Button(btn_frame, text="삭제", command=delete_item).pack(side='left', padx=5, expand=True)
        
        if config_key == 'materials':
            ttk.Button(btn_frame, text="📦 재고 일괄 등록", command=bulk_register_to_stock).pack(side='left', padx=5, expand=True)
            
        ttk.Button(btn_frame, text="닫기", command=on_close).pack(side='left', padx=5, expand=True)

    def bulk_register_materials_to_stock(self, material_list):
        """Register items from the preferred list to the main Material inventory if they don't exist or are inactive"""
        print(f"DEBUG: Starting bulk registration for {len(material_list)} items...")
        if not material_list:
            messagebox.showwarning("경고", "등록할 품목이 목록에 없습니다.")
            return
            
        added_count = 0
        reactivated_count = 0
        existing_indices = {} # Name.lower() -> index in materials_df
        max_id = 10000
        
        # Ensure materials_df is loaded
        if self.materials_df is None:
            self.materials_df = pd.DataFrame(columns=['MaterialID', '품목명', 'Active'])

        if not self.materials_df.empty:
            # Collect existing names and their positions
            # [STABILITY] Build index using (Name, Model) to allow parent items even if sub-items exist
            existing_keys = {} # (name_lower, model_lower) -> index
            for idx, mat in self.materials_df.iterrows():
                nm = str(mat.get('품목명', '')).strip().lower()
                md = str(mat.get('모델명', '')).strip().lower()
                if nm:
                    existing_keys[(nm, md)] = idx
            
            try:
                # Calculate next MaterialID
                if 'MaterialID' in self.materials_df.columns:
                    valid_ids = pd.to_numeric(self.materials_df['MaterialID'], errors='coerce').dropna()
                    if not valid_ids.empty:
                        max_id = int(valid_ids.max())
            except Exception as e:
                print(f"DEBUG: Error calculating max_id: {e}")
                max_id = 10000

        new_rows = []
        for mat in material_list:
            mat_clean = str(mat).strip()
            if not mat_clean: continue
            
            mat_lower = mat_clean.lower()
            # Check for parent item (empty model)
            if (mat_lower, '') in existing_keys:
                idx = existing_keys[(mat_lower, '')]
                # Check if it's inactive
                curr_active = self.materials_df.at[idx, 'Active']
                if str(curr_active) == '0' or curr_active == 0:
                    # Reactivate!
                    self.materials_df.at[idx, 'Active'] = 1
                    reactivated_count += 1
                    print(f"DEBUG: Reactivated item: {mat_clean}")
                else:
                    print(f"DEBUG: Skipping '{mat_clean}' (Already Active)")
            else:
                # Add as New
                max_id += 1
                new_row = {
                    'MaterialID': max_id,
                    '품목명': mat_clean,
                    'SN': '',
                    '모델명': '',
                    '관리단위': '매',
                    '수량': 0,
                    '재고하한': 0,
                    'Active': 1,
                    '구분': '소모품'
                }
                for col in self.materials_df.columns:
                    if col not in new_row:
                        new_row[col] = ''
                
                new_rows.append(new_row)
                added_count += 1
                print(f"DEBUG: Prepared new item: {mat_clean} (ID: {max_id})")
        
        if new_rows or reactivated_count > 0:
            try:
                if new_rows:
                    new_df = pd.DataFrame(new_rows)
                    new_df = new_df.reindex(columns=self.materials_df.columns, fill_value='')
                    self.materials_df = pd.concat([self.materials_df, new_df], ignore_index=True)
                
                # Save and Refresh
                self.save_data()
                self.update_stock_view()
                self.update_material_combo()
                
                msg = ""
                if added_count > 0: msg += f"{added_count}개의 품목 신규 등록 완료\n"
                if reactivated_count > 0: msg += f"{reactivated_count}개의 품목 재활성화(복구) 완료"
                
                messagebox.showinfo("등록 완료", msg)
            except Exception as e:
                print(f"DEBUG: Error during bulk registration: {e}")
                messagebox.showerror("오류", f"일괄 등록 중 오류가 발생했습니다:\n{e}")
        else:
            messagebox.showinfo("정보", "이미 모든 품목이 재고 현황에 활성화되어 있습니다.")

    def refresh_ui_for_list_change(self, config_key):
        """Update all related UI elements after a list (sites, users, etc) has changed"""
        # Dictionary mapping config keys to their current values
        list_map = {
            'sites': self.sites,
            'users': self.users,
            'warehouses': self.warehouses,
            'equipments': self.equipments,
            'worktimes': self.worktimes,
            'vehicles': self.vehicles,
            'companies': getattr(self, 'companies', []),
            'daily_units': getattr(self, 'daily_units', [])
        }
        
        if config_key not in list_map:
            return
            
        current_vals = list_map[config_key]
        # [NEW] Ensure uniqueness and strip whitespace before sorting
        unique_vals = sorted(list(set([str(v).strip() for v in current_vals if v])))
        sorted_vals = unique_vals
        
        # [FIX] Update the list in-place to ensure all UI dialogs holding references see the changes
        if isinstance(current_vals, list):
            current_vals[:] = sorted_vals
        
        # 1. Update standard widgets
        if config_key == 'sites':
            if hasattr(self, 'cb_daily_site'): self.cb_daily_site['values'] = sorted_vals
            if hasattr(self, 'cb_trans_site'): self.cb_trans_site['values'] = sorted_vals
            if hasattr(self, 'cb_daily_filter_site'):
                self.cb_daily_filter_site['values'] = ['전체'] + sorted_vals
            if hasattr(self, 'cb_budget_site'): self.cb_budget_site['values'] = sorted_vals
            if hasattr(self, 'cb_budget_view_site'): self.cb_budget_view_site['values'] = sorted_vals
        elif config_key == 'companies':
            # [NEW] Update all company comboboxes in daily usage
            if hasattr(self, 'cb_daily_company'): self.cb_daily_company['values'] = sorted_vals
            if hasattr(self, 'cb_daily_filter_company'):
                self.cb_daily_filter_company['values'] = ['전체'] + sorted_vals
        elif config_key == 'daily_units':
            print(f"DEBUG: Refreshing All Unit Comboboxes with: {sorted_vals}")
            # [FIX] daily_units is already updated in-place above via current_vals[:] = sorted_vals
            # 1. Re-merge to include DB units and update all UI automatically
            self.update_registration_combos()
            # update_registration_combos() handles cb_daily_unit and cb_unit automatically
        elif config_key == 'users':
            # Updated to match current attribute names if needed, 
            # but usually it's cb_daily_user for 1, and cb_daily_user{i} for 2-10
            for i in range(1, 11):
                attr = 'cb_daily_user' if i == 1 else f'cb_daily_user{i}'
                if hasattr(self, attr):
                    widget = getattr(self, attr)
                    if isinstance(widget, WorkerCompositeWidget):
                         widget.cb_name['values'] = sorted_vals
                    elif hasattr(widget, 'configure'):
                        try: widget['values'] = sorted_vals
                        except: pass
            
            if hasattr(self, 'ent_user'): self.ent_user['values'] = sorted_vals
        elif config_key == 'worktimes':
            # Update all worktime fields (1-10)
            for i in range(1, 11):
                attr = f'ent_worktime{i}'
                if hasattr(self, attr):
                    widget = getattr(self, attr)
                    if hasattr(widget, 'configure'):
                        try: 
                            widget['values'] = sorted_vals
                        except: pass
        elif config_key == 'equipments' or config_key == 'materials':
            # Both affect the unified material/equipment suggestion lists
            self.update_material_combo()
        elif config_key == 'test_items' or config_key == 'applied_codes':
            # These are handled by dynamic lambdas in _bind_combobox_word_suggest
            pass 
        elif config_key == 'warehouses':
            if hasattr(self, 'cb_warehouse'): self.cb_warehouse['values'] = sorted_vals
        elif config_key == 'vehicles':
            if hasattr(self, 'vehicle_boxes'):
                for box in self.vehicle_boxes:
                    box.update_vehicle_list(sorted_vals)
            for key, widget_instance in self.vehicle_inspections.items():
                widget_instance.update_vehicle_list(sorted_vals)
        elif config_key == 'co_code' or config_key == '회사코드':
            # [NEW] Update company code combobox in daily usage
            if hasattr(self, 'cb_daily_co_code'): self.cb_daily_co_code['values'] = sorted_vals

        # Always save configuration after any list change
        self.save_tab_config()


        # 2. Update ALL draggable widgets (clones) that depend on this list
        for key, container in self.draggable_items.items():
            # Heuristic: if manage_list_key is missing but label suggests it's a worker/user/site
            m_key = getattr(container, '_manage_list_key', None)
            if not m_key:
                if hasattr(container, '_label_widget'):
                    lbl_text = container._label_widget.cget('text').lower()
                    if config_key == 'users' and any(x in lbl_text for x in ['작업자', '담당자', 'user', 'worker']):
                        m_key = 'users'
                        container._manage_list_key = 'users'
                    elif config_key == 'sites' and any(x in lbl_text for x in ['현장', 'site']):
                        m_key = 'sites'
                        container._manage_list_key = 'sites'
                    elif config_key == 'equipments' and any(x in lbl_text for x in ['장비', 'equip']):
                        m_key = 'equipments'
                        container._manage_list_key = 'equipments'
                    elif config_key == 'vehicles' and any(x in lbl_text for x in ['차량', 'vehicle']):
                        m_key = 'vehicles'
                        container._manage_list_key = 'vehicles'

            if m_key == config_key:
                if hasattr(container, '_widget') and hasattr(container._widget, 'config'):
                    try:
                        container._widget['values'] = sorted_vals
                    except:
                        pass

    def remove_box(self, key):
        """Intelligently remove a box: hide standard ones, destroy custom ones"""
        is_dynamic = key.startswith('memo_') or key.startswith('clone_') or key.startswith('checklist_') or key.startswith('vehicle_inspection_')
        
        # Allow deletion of dynamic widgets even if layout is locked
        if self.layout_locked and not is_dynamic: return
        
        if is_dynamic:
            # Permanent deletion for custom/dynamic items
            self.destroy_custom_widget(key)
        else:
            # Hiding for standard items (can be restored via Reset All)
            widget = self.draggable_items.get(key)
            if widget:
                # We reuse the existing hide_widget logic
                self.hide_widget(None, widget=widget)

    def destroy_custom_widget(self, key):
        """Destroy and remove from config any dynamic widget (clone or memo)"""
        # Allow destruction even if layout_locked (since it's a content management action)
        
        widget = self.draggable_items.get(key)
        if widget:
            widget.destroy()
            if key in self.draggable_items:
                del self.draggable_items[key]
            if key in self.memos:
                del self.memos[key]
            if key in self.checklists:
                del self.checklists[key]
            if key in self.vehicle_inspections:
                widget_to_remove = self.vehicle_inspections[key]
                if widget_to_remove in self.vehicle_boxes:
                    self.vehicle_boxes.remove(widget_to_remove)
                del self.vehicle_inspections[key]

            
            # Clean from in-memory config to prevent resurrection on next save_tab_config call
            if hasattr(self, 'tab_config') and 'draggable_geometries' in self.tab_config:
                if key in self.tab_config['draggable_geometries']:
                    del self.tab_config['draggable_geometries'][key]
            
            # Clean from config immediately to prevent resurrection on next load
            try:
                import json
                if os.path.exists(self.config_path):
                    with open(self.config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                    
                    if 'draggable_geometries' in config and key in config['draggable_geometries']:
                        del config['draggable_geometries'][key]
                        
                    with open(self.config_path, 'w', encoding='utf-8') as f:
                        json.dump(config, f, ensure_ascii=False, indent=2)
            except:
                pass
            
            self.save_tab_config()

    def rename_widget_label(self, key):
        """Show a simple dialog to rename a widget's label"""
        if self.layout_locked: return
        
        widget = self.draggable_items.get(key)
        if not widget: return
        
        current_text = ""
        if hasattr(widget, '_label_widget'):
            current_text = widget._label_widget.cget('text')
        elif key in self.memos:
            current_text = self.memos[key]['title_entry'].get()
        elif key in self.checklists:
            current_text = self.checklists[key]['title_entry'].get()
            
        new_name = simpledialog.askstring("이름 변경", "새 이름을 입력하세요:", initialvalue=current_text)
        if new_name is not None:
            if hasattr(widget, '_label_widget'):
                widget._label_widget.config(text=new_name)
            elif key in self.memos:
                self.memos[key]['title_entry'].delete(0, 'end')
                self.memos[key]['title_entry'].insert(0, new_name)
            elif key in self.checklists:
                self.checklists[key]['title_entry'].delete(0, 'end')
                self.checklists[key]['title_entry'].insert(0, new_name)
            self.save_tab_config()

    def clone_widget(self, key):
        """Create a clone of an existing widget as a new custom box"""
        if self.layout_locked: return
        
        orig = self.draggable_items.get(key)
        if not orig: return
        
        import time
        new_key = f"clone_{int(time.time() * 1000)}"
        
        label_text = ""
        if hasattr(orig, '_label_widget'):
            label_text = orig._label_widget.cget('text')
        
        if hasattr(orig, '_widget_class'):
            # It's a container created via create_draggable_container
            cont, w = self.create_draggable_container(
                self.entry_inner_frame, 
                label_text, 
                orig._widget_class, 
                new_key, 
                manage_list_key=getattr(orig, '_manage_list_key', None), # Pass manage_list_key
                **orig._widget_kwargs
            )
            
            # Copy value from original widget
            if hasattr(orig, '_widget'):
                try:
                    current_val = str(orig._widget.get()) # Ensure string
                    
                    # Try generic Entry-like setting (works for Entry and Combobox text area)
                    if hasattr(w, 'delete') and hasattr(w, 'insert'):
                        try:
                            w.delete(0, 'end')
                            w.insert(0, current_val)
                        except:
                            # Readonly comboboxes might fail delete/insert
                            pass
                            
                    # Try specific set method (Combobox, Scale, etc)
                    if hasattr(w, 'set'):
                        w.set(current_val)
                        
                except Exception as e:
                    print(f"Failed to copy value: {e}")
            
            cont.place(x=50, y=50) # Start position
            self.save_tab_config()

        elif key in self.memos:
            # It's a memo
            content = self.memos[key]['text_widget'].get('1.0', 'end-1c')
            title = self.memos[key]['title_entry'].get()
            self.add_new_memo(initial_text=content, initial_title=title, key=new_key)
            self.save_tab_config()
        elif key in self.checklists:
            # It's a checklist
            self.duplicate_checklist(key)
            self.save_tab_config()

    def _bind_recursive(self, widget, target_container):
        """Recursively bind drag events to widget and its children"""
        from functools import partial
        
        # Bind events to this widget, targeting the container
        # Note: We use add=True to avoid overwriting existing bindings if possible, 
        # but for drag we usually want exclusive control or at least priority.
        # Here we just bind standard.
        
        # We need to capture the target_container in the callback
        widget.bind("<Button-3>", partial(self.on_drag_start, widget=target_container))
        widget.bind("<Shift-Button-3>", partial(self.on_resize_start, widget=target_container))
        widget.bind("<B3-Motion>", partial(self.on_mouse_motion, widget=target_container))
        widget.bind("<Double-Button-3>", partial(self.reset_widget_position, widget=target_container))
        widget.bind("<Control-Button-3>", partial(self.hide_widget, widget=target_container))
        widget.bind("<ButtonRelease-3>", partial(self.on_drag_stop, widget=target_container))
        
        # Recurse for children
        try:
            for child in widget.winfo_children():
                # Toplevel(DateEntry 달력 팝업 등) 내부는 절대 순회하지 않음
                # → winfo_children() 재귀만으로도 Toplevel이 화면에 나타날 수 있음
                if child.winfo_class() == 'Toplevel':
                    continue
                self._bind_recursive(child, target_container)
        except:
            pass

    def hide_widget(self, event, widget=None):
        """Hide a widget (Ctrl + Right Click)"""
        if self.layout_locked:
            return "break"
            
        if widget is None:
            widget = event.widget
            
        # [STABILITY FIX] Prevent hiding core widgets
        if hasattr(widget, '_config_key') and widget._config_key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
            messagebox.showwarning("보호됨", f"'{widget._config_key}' 항목은 화면의 필수 구성 요소이므로 숨길 수 없습니다.")
            return "break"
        
        # Remove from place layout
        if widget.winfo_manager() == 'place':
            widget.place_forget()
        elif widget.winfo_manager() == 'grid':
            widget.grid_forget()
            
        # Remove placeholder if exists
        self._remove_placeholder(widget)
        
        # Mark as hidden in config
        if hasattr(widget, '_config_key') and widget._config_key:
            try:
                import json
                if os.path.exists(self.config_path):
                    with open(self.config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                else:
                    config = {}
                
                if 'draggable_geometries' not in config:
                    config['draggable_geometries'] = {}
                
                if widget._config_key not in config['draggable_geometries']:
                    config['draggable_geometries'][widget._config_key] = {}
                
                config['draggable_geometries'][widget._config_key]['hidden'] = True
                
                with open(self.config_path, 'w', encoding='utf-8') as f:
                    json.dump(config, f, ensure_ascii=False, indent=2)
            except Exception as e:
                print(f"Failed to save hide status: {e}")
        return "break"

    def make_draggable(self, widget, config_key=None):
        """Make a widget draggable and resizable with right mouse button"""
        widget._config_key = config_key
        # Save original grid info for reset/placeholder
        # We need to do this immediately while it's still in the grid
        widget._original_grid_info = widget.grid_info()
        
        # Recursively bind to ensure clicking anywhere works
        self._bind_recursive(widget, widget)

    def make_header_draggable(self, widget, target_container):
        """Make a specific widget (header/label) draggable with Left Mouse Button targeting a container"""
        from functools import partial
        widget.bind("<Button-1>", partial(self.on_drag_start, widget=target_container))
        widget.bind("<B1-Motion>", partial(self.on_mouse_motion, widget=target_container))
        widget.bind("<ButtonRelease-1>", partial(self.on_drag_stop, widget=target_container))
        
    def reset_widget_position(self, event, widget=None):
        """Reset widget to original grid position"""
        if self.layout_locked:
            return "break"
            
        if widget is None:
            widget = event.widget
        
        # Remove from place layout
        widget.place_forget()
        
        # Remove placeholder if exists
        self._remove_placeholder(widget)
            
        # Restore to grid
        if hasattr(widget, '_original_grid_info'):
            widget.grid(**widget._original_grid_info)
        
        # Reset size variables if any
        if hasattr(widget, '_start_width'): del widget._start_width
        if hasattr(widget, '_start_height'): del widget._start_height
        
        # Remove from config
        if hasattr(widget, '_config_key') and widget._config_key:
            try:
                import json
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                
                if 'draggable_geometries' in config and widget._config_key in config['draggable_geometries']:
                    del config['draggable_geometries'][widget._config_key]
                
                # Backward compatibility: also check top level
                if widget._config_key in config:
                    del config[widget._config_key]
                    
                with open(self.config_path, 'w', encoding='utf-8') as f:
                    json.dump(config, f, ensure_ascii=False, indent=2)
                    
            except Exception as e:
                print(f"Failed to reset config: {e}")

    def reset_all_widgets_layout(self):
        """Reset all widgets to their original grid slots, unhide them, and reset paned window sashes"""
        if messagebox.askyesno("초기화", "모든 항목의 위치와 크기를 초기화하고 숨겨진 항목을 다시 표시하시겠습니까?\n(창 분할 위치도 초기화됩니다.)"):
            # 1. Reset draggable widgets
            for key, widget in list(self.draggable_items.items()):
                # Explicitly unhide if it was hidden
                if widget.winfo_manager() == '':
                    # For grid, we just grid it back
                    if hasattr(widget, '_original_grid_info'):
                         widget.grid(**widget._original_grid_info)
                
                self.reset_widget_position(None, widget=widget)
            
            # 1.05 [NEW] Use custom defaults if they exist
            if os.path.exists(self.config_path):
                try:
                    with open(self.config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                    
                    custom_defaults = config.get('custom_default_geometries', {})
                    if custom_defaults:
                        # Clear existing geometries to ensure fresh start
                        if 'draggable_geometries' not in config: config['draggable_geometries'] = {}
                        
                        for key, geo in custom_defaults.items():
                            if key in self.draggable_items:
                                w = self.draggable_items[key]
                                w.place(x=geo['x'], y=geo['y'], width=geo['width'], height=geo['height'])
                                # Also update config so it persists as the new current layout
                                config['draggable_geometries'][key] = geo
                        
                        # Save the updated config back to disk and memory
                        with open(self.config_path, 'w', encoding='utf-8') as f:
                            json.dump(config, f, ensure_ascii=False, indent=2)
                        self.tab_config = config
                except Exception as e:
                    print(f"Error applying custom defaults: {e}")
                
            # 1.1 Restore parent propagation safely and reset size
            if hasattr(self, 'entry_inner_frame'):
                self.entry_inner_frame.pack_propagate(True)
                self.entry_inner_frame.grid_propagate(True)
                self.entry_inner_frame.config(height=0, width=0) # Let it auto-size for reset
                
                # Perform a layout update pass
                self.root.update_idletasks()
                
                # Re-apply stability locks
                self._adjust_parent_height(self.entry_inner_frame, force=True)
                self.entry_inner_frame.pack_propagate(False)
                self.entry_inner_frame.grid_propagate(False)
                
                # Refresh scrollregion if in canvas
                if hasattr(self, 'entry_canvas'):
                    self.entry_canvas.configure(scrollregion=self.entry_canvas.bbox("all"))
            
            # 2. Reset sash positions (splitters)
            try:
                if hasattr(self, 'daily_usage_paned'):
                    # Give more space to the entry form by default (500px)
                    self.daily_usage_paned.sashpos(0, 500) 
                if hasattr(self, 'daily_history_paned'):
                    total_w = self.daily_history_paned.winfo_width()
                    if total_w > 100:
                        self.daily_history_paned.sashpos(0, int(total_w * 0.7))
                    else:
                        self.daily_history_paned.sashpos(0, 800)
            except:
                pass

            self.save_tab_config()
            messagebox.showinfo("완료", "레이아웃과 창 분할 위치가 초기화되었습니다.")

    def save_current_layout_as_default(self):
        """Save the current layout as the custom 'reset' default"""
        try:
            current_geos = {}
            for key, widget in self.draggable_items.items():
                # [FIX] Capture all managed widgets (grid or place) 
                # This ensures the snapshot is complete even if some items weren't moved.
                manager = widget.winfo_manager()
                if manager in ['place', 'grid']:
                    current_geos[key] = {
                        'x': widget.winfo_x(),
                        'y': widget.winfo_y(),
                        'width': widget.winfo_width(),
                        'height': widget.winfo_height(),
                        'hidden': False
                    }
            
            if not current_geos:
                messagebox.showwarning("실패", "수동으로 배치된(드래그된) 항목이 없어 기본값으로 저장할 수 없습니다.")
                return

            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            else:
                config = {}
                
            config['custom_default_geometries'] = current_geos
            
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
                
            messagebox.showinfo("완료", "현재 레이아웃이 '기본 배치'로 저장되었습니다.\n앞으로 [초기화] 시 이 배치로 되돌아갑니다.")
        except Exception as e:
            messagebox.showerror("오류", f"기본 배치 저장 중 오류 발생: {e}")
            
    def toggle_layout_lock(self):
        """Toggle layout locking state"""
        self.layout_locked = not self.layout_locked
        if hasattr(self, 'btn_lock_layout'):
            if self.layout_locked:
                self.btn_lock_layout.config(text="🔒 배치 고정됨")
                self.style.configure("Lock.TButton", foreground="black")
            else:
                self.btn_lock_layout.config(text="🔓 배치 수정 중")
                self.style.configure("Lock.TButton", foreground="red")
        
        # Save layout lock state to config immediately
        if not hasattr(self, 'tab_config'):
            self.tab_config = {}
        self.tab_config['layout_locked'] = self.layout_locked
        
        # [FIX] Unlocking is no longer destructive. 
        # It simply enables movement without resetting positions.
        if self.layout_locked:
            print("Layout locked - current positions preserved")
        else:
            print("Layout unlocked - movement enabled")
        
        # Force save the state immediately to file
        self.save_tab_config(force=True)
        
        print(f"Layout lock {'enabled' if self.layout_locked else 'disabled'}")

    def on_drag_stop(self, event, widget=None):
        """Handle end of dragging or resizing and auto-save"""
        if widget is None:
            widget = event.widget
        if hasattr(widget, '_interaction_mode'):
            mode = getattr(widget, '_interaction_mode')
            del widget._interaction_mode
            
            # Update parent height if something moved or resized
            self._adjust_parent_height(widget.master, force=True)
            
            # Auto-save layout
            self.save_tab_config()


    def add_new_memo(self, initial_text="", initial_title="메모", key=None):
        """Create a new movable/editable/copyable memo box with editable title"""
        import time
        if key is None:
            key = f"memo_{int(time.time() * 1000)}"
        
        # Create container

        memo_container = ttk.LabelFrame(self.entry_inner_frame)
        
        # Controls frame (top of memo)
        ctrl_frame = ttk.Frame(memo_container)
        ctrl_frame.pack(fill='x', side='top')
        
        # Editable Title Entry
        title_entry = ttk.Entry(ctrl_frame, font=('Arial', 9, 'bold'), width=15)
        title_entry.pack(side='left', padx=2)
        title_entry.insert(0, initial_title)
        title_entry.bind('<FocusOut>', lambda e: self.save_tab_config())
        
        btn_copy = ttk.Button(ctrl_frame, text="📋", width=3, command=lambda: self.duplicate_memo(key))
        btn_copy.pack(side='right')
        
        btn_del = ttk.Button(ctrl_frame, text="❌", width=3, command=lambda: self.remove_box(key))
        btn_del.pack(side='right')
        
        # Text area
        text_area = tk.Text(memo_container, wrap='word', height=5, width=30, font=('Arial', 10), bg=self.theme_bg, highlightthickness=0)
        text_area.pack(fill='both', expand=True, padx=2, pady=2)
        text_area.insert('1.0', initial_text)
        
        # Bind text change to auto-save
        text_area.bind('<FocusOut>', lambda e: self.save_tab_config())
        
        # Make draggable (Right Click)
        self.make_draggable(memo_container, key)
        
        # Make header draggable (Left Click)
        self.make_header_draggable(ctrl_frame, memo_container)
        self.draggable_items[key] = memo_container
        self.memos[key] = {
            'container': memo_container, 
            'text_widget': text_area,
            'title_entry': title_entry
        }
        
        # Initial placement if not loaded from config (will be handled by load_tab_config if exists)
        if key not in getattr(self, '_loading_memos', []):
            memo_container.place(x=1450, y=50) # Default start position
        
        return memo_container

    def duplicate_memo(self, key):
        """Duplicate an existing memo with its title and content"""
        if self.layout_locked: return
        if key in self.memos:
            content = self.memos[key]['text_widget'].get('1.0', 'end-1c')
            title = self.memos[key]['title_entry'].get()
            self.add_new_memo(initial_text=content, initial_title=title)
            self.save_tab_config()

    def add_new_checklist(self, initial_data=None, initial_title="체크리스트", key=None):
        """Create a new movable/editable checklist box"""
        import time
        if key is None:
            key = f"checklist_{int(time.time() * 1000)}"
        
        # Create container
        check_container = ttk.LabelFrame(self.entry_inner_frame)
        
        # Controls frame (top of checklist)
        ctrl_frame = ttk.Frame(check_container)
        ctrl_frame.pack(fill='x', side='top')
        
        # Editable Title Entry
        title_entry = ttk.Entry(ctrl_frame, font=('Arial', 9, 'bold'), width=15)
        title_entry.pack(side='left', padx=2)
        title_entry.insert(0, initial_title)
        title_entry.bind('<FocusOut>', lambda e: self.save_tab_config())
        
        btn_copy = ttk.Button(ctrl_frame, text="📋", width=3, command=lambda: self.duplicate_checklist(key))
        btn_copy.pack(side='right')
        
        btn_del = ttk.Button(ctrl_frame, text="❌", width=3, command=lambda: self.remove_box(key))
        btn_del.pack(side='right')
        
        # Add Item Area
        add_frame = ttk.Frame(check_container)
        add_frame.pack(fill='x', padx=2, pady=2)
        
        new_item_var = tk.StringVar()
        entry_new = ttk.Entry(add_frame, textvariable=new_item_var, width=20)
        entry_new.pack(side='left', fill='x', expand=True)
        
        def add_item(event=None):
             text = new_item_var.get().strip()
             if text:
                 self.add_checklist_item(item_frame, text, False, key)
                 new_item_var.set("")
                 self.save_tab_config()
        
        def _adjust_parent_height(self, parent, force=False):
            """Adjust parent frame height with performance check"""
            try:
                # Only update idletasks if forced or we're not in the middle of a high-speed interaction
                # This is the single biggest cause of UI stutter.
                if force:
                    parent.update_idletasks()
            except:
                pass
        
        entry_new.bind('<Return>', add_item)
        btn_add = ttk.Button(add_frame, text="➕", width=3, command=add_item)
        btn_add.pack(side='right')

        # Scrollable Frame for Items
        canvas_frame = ttk.Frame(check_container)
        canvas_frame.pack(fill='both', expand=True, padx=2, pady=2)
        
        canvas = tk.Canvas(canvas_frame, height=100, width=200, bg=self.theme_bg, highlightthickness=0) # Match theme
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        item_frame = ttk.Frame(canvas)
        
        item_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=item_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Scroll binding is handled globally in MaterialManager.__init__

        # Make draggable
        self.make_draggable(check_container, key)
        self.draggable_items[key] = check_container
        self.checklists[key] = {
            'container': check_container,
            'title_entry': title_entry,
            'item_frame': item_frame,
            'items': [] # List of item widgets/vars will be managed dynamically via children of item_frame
        }
        
        # Add initial items if provided
        if initial_data:
            for item in initial_data:
                self.add_checklist_item(item_frame, item.get('text', ''), item.get('checked', False), key)
        
        # Initial placement
        if key not in getattr(self, '_loading_memos', []): # Reuse loading flag or logic
            check_container.place(x=1450, y=50)
            
        return check_container

    def save_single_vehicle_data(self, widget):
        """Save ONLY the requested vehicle's data as a standalone entry in the daily log"""
        try:
            # 1. Collect minimal context
            usage_date = str(self.ent_daily_date.get_date())
            site = self.cb_daily_site.get().strip()
            if not site:
                messagebox.showwarning("입력 필요", "차량 점검을 기록할 '현장'을 선택해주세요.")
                return

            # 2. Collect vehicle data
            v_data = widget.get_data()
            v_no = v_data.get('vehicle_info', '').strip()
            if not v_no:
                messagebox.showwarning("입력 필요", "차량번호를 입력해주세요.")
                return

            # Confirm with user
            if not messagebox.askyesno("확인", f"[{v_no}] 차량 점검 정보만 단독으로 저장하시겠습니까?"):
                return

            # 3. Create record (Everything else is zero/empty)
            reserved = ['vehicle_info', 'mileage', 'remarks', '_raw_mileage']
            checks = "|".join([f"{k}:{v}" for k, v in v_data.items() if k not in reserved and v])

            record = {
                'Date': usage_date,
                '업체명': self.cb_daily_company.get().strip() or "현장기록",
                'Site': site,
                'User': "차량점검", # Visual tag in worker column
                'WorkTime': "",
                '장비명': self.cb_daily_equip.get().strip(),
                '검사방법': self.cb_daily_test_method.get().strip(),
                '회사코드': "",
                'Usage': 0.0, '단가': 0.0, '출장비': 0.0, '검사비': 0.0,
                '검사량': 0.0,
                'OT': 0.0, 'OT금액': 0.0,
                'MaterialID': "차량점검",
                '차량번호': v_no,
                '주행거리': v_data.get('_raw_mileage', '0'),
                '차량점검': checks,
                '차량비고': v_data.get('remarks', ''),
                'EntryTime': datetime.datetime.now(),
                '(Full작업자)': "차량점검"
            }
            # Initialize RTK/NDT columns as 0
            for rtk_cat in ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타", "RTK총계"]:
                record[f'RTK_{rtk_cat}'] = 0
            for ndt_cat in ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]:
                record[f'NDT_{ndt_cat}'] = 0

            # 4. Append and Save
            new_df = pd.DataFrame([record])
            self.daily_usage_df = pd.concat([self.daily_usage_df, new_df], ignore_index=True)
            self.save_data()
            
            # 5. UI Updates
            self.update_daily_usage_view()
            self.refresh_inquiry_filters()
            
            messagebox.showinfo("성공", f"[{v_no}] 차량 정보가 기록에 추가되었습니다.\n(실적 없이 차량 단독 기록으로 저장됨)")
            
            # [V12_QUICK_MULTI_VEHICLE] Reset only vehicle-specific fields for next entry
            if hasattr(widget, 'reset_fields'):
                widget.reset_fields()
            
        except Exception as e:
            messagebox.showerror("오류", f"차량 정보 저장 중 오류 발생: {e}")

    def add_vehicle_inspection_box(self, initial_data=None, key=None):
        """Create a new movable vehicle inspection box"""
        import time
        if key is None:
            key = f"vehicle_inspection_{int(time.time() * 1000)}"
            
        # Create draggable container
        container, widget = self.create_draggable_container(
            self.entry_inner_frame, "차량 점검", VehicleInspectionWidget, key,
            manage_list_key='vehicles',
            theme_bg=getattr(self, 'theme_bg', '#f0f0f0'),
            vehicle_list=getattr(self, 'vehicles', []),
            on_save=self.save_single_vehicle_data
        )
        
        # Track specifically for data gathering
        self.vehicle_inspections[key] = widget
        if widget not in self.vehicle_boxes:
            self.vehicle_boxes.append(widget)
        
        # Load initial data if provided
        if initial_data:
            widget.set_data(initial_data)
            
        # Initial placement if new (Below the trigger button in the main entry frame)
        if key not in getattr(self, '_loading_memos', []):
            container.place(x=1450, y=50)
            
        return container

    def add_checklist_item(self, parent_frame, text, checked, checklist_key):
        """Add a single item row to the checklist"""
        row_frame = ttk.Frame(parent_frame)
        row_frame.pack(fill='x', pady=1)
        
        var = tk.BooleanVar(value=checked)
        cb = ttk.Checkbutton(row_frame, variable=var, command=lambda: self.save_tab_config())
        cb.pack(side='left')
        
        entry = ttk.Entry(row_frame)
        entry.insert(0, text)
        entry.pack(side='left', fill='x', expand=True)
        entry.bind('<FocusOut>', lambda e: self.save_tab_config())
        
        def delete_this_item():
            row_frame.destroy()
            self.save_tab_config()
            
        btn_del_item = ttk.Label(row_frame, text="❌", font=('Arial', 8), cursor='hand2', foreground='gray')
        btn_del_item.pack(side='right', padx=2)
        btn_del_item.bind('<Button-1>', lambda e: delete_this_item())
        
        # Store refs in widget for retrieval during save
        row_frame._checklist_data = {'var': var, 'entry': entry}

    def duplicate_checklist(self, key):
        """Duplicate an existing checklist"""
        if self.layout_locked: return
        if key in self.checklists:
            # get current data
            original_items = []
            item_frame = self.checklists[key]['item_frame']
            for child in item_frame.winfo_children():
                if hasattr(child, '_checklist_data'):
                    data = child._checklist_data
                    original_items.append({
                        'text': data['entry'].get(),
                        'checked': data['var'].get()
                    })
            
            title = self.checklists[key]['title_entry'].get()
            self.add_new_checklist(initial_data=original_items, initial_title=title)
            self.save_tab_config()


    def on_drag_start(self, event, widget=None):
        """Begin dragging widget"""
        if self.layout_locked:
            return "break" # Prevent movement and stop propagation
            
        if widget is None:
            widget = event.widget
        widget._interaction_mode = 'move'
        
        # Save absolute start position of mouse
        widget._drag_start_root_x = event.x_root
        widget._drag_start_root_y = event.y_root
        
        # Save initial widget position relative to parent
        widget._drag_start_pos_x = widget.winfo_x()
        widget._drag_start_pos_y = widget.winfo_y()
        
        # Ensure we have grid info (redundant but safe)
        if not hasattr(widget, '_original_grid_info') and widget.grid_info():
            widget._original_grid_info = widget.grid_info()
        
    def on_resize_start(self, event, widget=None):
        """Begin resizing widget"""
        if self.layout_locked:
            return "break"
            
        if widget is None:
            widget = event.widget
        widget._interaction_mode = 'resize'
        
        # Save absolute start position of mouse
        widget._drag_start_root_x = event.x_root
        widget._drag_start_root_y = event.y_root
        
        # Save initial size
        widget._start_width = widget.winfo_width()
        widget._start_height = widget.winfo_height()
        
        # Ensure we have grid info
        if not hasattr(widget, '_original_grid_info') and widget.grid_info():
            widget._original_grid_info = widget.grid_info()
        return "break"
        
    def on_mouse_motion(self, event, widget=None):
        """Handle dragging or resizing motion with performance throttling"""
        if self.layout_locked:
            return "break"
            
        if widget is None:
            widget = event.widget
        
        if not hasattr(widget, '_interaction_mode'):
            return
        
        # PERFORMANCE THROTTLE: Limit updates to ~60fps (16ms)
        curr_time = time.time()
        if curr_time - self._last_motion_time < 0.016:
            # Still update the physical position of the widget being interacted with
            # or the user will feel lag in the initial drag/resize itself.
            self._update_widget_position(event, widget)
            return
        
        self._last_motion_time = curr_time
        
        # Apply positioning only (No collision, No auto-resize)
        self._update_widget_position(event, widget)

    def _update_widget_position(self, event, widget):
        """Internal helper to calculate and set widget position/size during interaction"""
        dx = event.x_root - widget._drag_start_root_x
        dy = event.y_root - widget._drag_start_root_y
        parent = widget.master
        parent_w = parent.winfo_width()
        parent_h = parent.winfo_height()
        
        if widget._interaction_mode == 'move':
            new_x = widget._drag_start_pos_x + dx
            new_y = widget._drag_start_pos_y + dy
            widget_w = widget.winfo_width()
            widget_h = widget.winfo_height()
            
            # Clamp
            if new_x < 0: new_x = 0
            elif new_x + widget_w > parent_w: new_x = max(0, parent_w - widget_w)
            
            # Loosen Y clamping: allow moving slightly below parent to trigger growth
            if new_y < 0: new_y = 0
            # Instead of strict clamp at parent_h, allow a small overflow to trigger resize logic
            elif new_y + widget_h > parent_h + 100: 
                new_y = parent_h + 100 - widget_h

            
            if widget.winfo_manager() != 'place':
                # Save original grid info and explicitly un-grid
                if not hasattr(widget, '_original_grid_info') and widget.grid_info():
                    widget._original_grid_info = widget.grid_info()
                self._ensure_placeholder(widget)
                widget.grid_forget() # Explicitly un-grid before placing
                widget.place(width=widget_w, height=widget_h)
                
            widget.place(x=new_x, y=new_y)
            widget.lift()
            
            # Auto-update scroll region if we're dragging near bounds
            self._ensure_canvas_scroll_region()
            
        elif widget._interaction_mode == 'resize':
            new_width = max(50, widget._start_width + dx)
            new_height = max(20, widget._start_height + dy)
            current_x = widget.winfo_x()
            current_y = widget.winfo_y()
            
            # Clamp
            if current_x + new_width > parent_w: new_width = max(50, parent_w - current_x)
            if current_y + new_height > parent_h: new_height = max(20, parent_h - current_y)
            
            if widget.winfo_manager() != 'place':
                # Save original grid info and explicitly un-grid
                if not hasattr(widget, '_original_grid_info') and widget.grid_info():
                    widget._original_grid_info = widget.grid_info()
                self._ensure_placeholder(widget)
                widget.grid_forget() # Explicitly un-grid before placing
                widget.place(x=current_x, y=current_y)
                
            widget.place(width=new_width, height=new_height)
            widget.lift()
        
        return "break"
        
    def _apply_push_down_logic(self, dragged_widget):
        """Recursively push widgets down if they overlap with the dragged widget"""
        try:
            # Current bounding box of the dragging widget
            x1 = dragged_widget.winfo_x()
            y1 = dragged_widget.winfo_y()
            w1 = dragged_widget.winfo_width()
            h1 = dragged_widget.winfo_height()
            
            padding = 10
            
            # Parent of the widgets
            parent = dragged_widget.master
            
            # OPTIMIZATION: Early filter candidates by parent to avoid massive iterations
            candidates = [w for w in self.draggable_items.values() if w.master == parent]
            
            for other in candidates:
                if other == dragged_widget:
                    continue
                
                # [STABILITY FIX] Skip if the 'other' widget's BOTTOM is ABOVE the dragged widget's TOP.
                # Strictly ignore anything physically above the current interaction zone.
                if (other.winfo_y() + other.winfo_height()) <= y1:
                    continue

                # Skip if not in the same parent
                if other.master != parent:
                    continue
                
                # Skip if hidden
                if not other.winfo_ismapped():
                    continue
                
                # Get current pos/size of the 'other' widget
                x2 = other.winfo_x()
                y2 = other.winfo_y()
                w2 = other.winfo_width()
                h2 = other.winfo_height()
                
                # 1. Horizontal Overlap Check
                if (x1 < x2 + w2) and (x1 + w1 > x2):
                    # 2. Vertical Collision Check
                    # If dragged_widget is hitting 'other' from the top
                    # We check if the bottom of dragged_widget is below the top of other
                    if y1 < y2 and (y1 + h1) > y2:
                        # New target Y for 'other'
                        new_y2 = y1 + h1 + padding
                        
                        # Only move if it's actually pushing it DOWN
                        if y2 < new_y2:
                            if other.winfo_manager() == 'grid':
                                # Convert grid to place
                                self._ensure_placeholder(other)
                                other.lift()
                                other.place(x=x2, y=new_y2, width=w2, height=h2)
                            else:
                                # Just update place
                                other.place(y=new_y2)
                            
                            # 3. Recursive Push
                            self._apply_push_down_logic(other)
        except Exception as e:
            # Silent fail to avoid interrupting drag
            print(f"Error in push-down logic: {e}")

    def _adjust_parent_height(self, parent, force=False):
        """Adjust parent frame height with performance check"""
        try:
            # Only update idletasks if forced or we're not in the middle of a high-speed interaction
            # This is the single biggest cause of UI stutter.
            if force:
                parent.update_idletasks()
            # 2. Start with the bounding box of all GRIDDED items
            try:
                # grid_bbox returns (x, y, width, height) of the grid
                bbox = parent.grid_bbox()
                required_h = bbox[1] + bbox[3] if bbox[3] > 0 else 0
                required_w = bbox[0] + bbox[2] if bbox[2] > 0 else 0
            except:
                required_h = 0
                required_w = 0

            # 3. Handle PLACE items as well (dragged/custom items)
            for child in parent.winfo_children():
                try:
                    manager = child.winfo_manager()
                    if manager == 'place':
                        info = child.place_info()
                        # Get relative or absolute y + height
                        y = child.winfo_y()
                        h = child.winfo_height()
                        x = child.winfo_x()
                        w = child.winfo_width()
                        required_h = max(required_h, y + h)
                        required_w = max(required_w, x + w)
                except:
                    continue
            
            # Add some padding
            new_height = required_h + 30
            
            # [STABILITY FIX] Guard against collapse but allow growth
            # Width is now dictated by canvas parent in responsive mode.
            if new_height < 500: new_height = 800
            
            # Only resize if the required dimensions are significantly different
            current_h = parent.winfo_height()
            if abs(current_h - new_height) > 10:
                 parent.config(height=new_height)
                 
                 # [RECURSIVE FIX] If this parent has a master that is part of the entry system,
                 # adjust its height too. 
                 if parent.master and parent.master != self.entry_inner_frame.master: # Don't go above canvas
                     self._adjust_parent_height(parent.master, force=force)
                     
        except Exception as e:
            print(f"Error adjusting parent height: {e}")


    def _ensure_placeholder(self, widget, width=None, height=None):
        """Ensure a placeholder exists in the grid where the widget used to be"""
        # Map widget to a placeholder attribute name dynamically
        # We can use the widget id or a dictionary, but simpler to attach to widget
        if not hasattr(widget, '_placeholder'):
            # Use provided dims or current widget dims
            w = width if width is not None else widget.winfo_width()
            h = height if height is not None else widget.winfo_height()
            
            # Create a frame to hold the space
            widget._placeholder = ttk.Frame(widget.master, width=w, height=h)
            
            # Grid it at the original position
            if hasattr(widget, '_original_grid_info'):
                widget._placeholder.grid(**widget._original_grid_info)
                # Ensure the placeholder doesn't shrink and holds its size
                widget._placeholder.grid_propagate(False)
                
    def _remove_placeholder(self, widget):
        """Remove placeholder for a widget"""
        if hasattr(widget, '_placeholder'):
            widget._placeholder.destroy()
            del widget._placeholder

        

    def setup_daily_usage_tab(self):
        """Setup the daily usage entry tab"""
        # Top frame for entry form
        # We use a canvas or large frame to allow free movement? 
        # Actually, we keep the entry_frame as the parent but use grid for initial layout
        # The user can then move them out of grid into place
        
        # Create PanedWindow for resizable frames
        self.daily_usage_paned = ttk.Panedwindow(self.tab_daily_usage, orient='vertical')
        self.daily_usage_paned.pack(fill='both', expand=True, padx=5, pady=5)  # Reduced padding
        
        # Save sash position on adjustment and lock it
        self.daily_usage_paned.bind("<ButtonRelease-1>", self._on_daily_usage_sash_changed)
        self.daily_usage_paned.bind("<Configure>", self._on_daily_usage_resize)
        # [FIX] Respect loaded config if available, otherwise default to False
        self.daily_usage_sash_locked = getattr(self, 'daily_usage_sash_locked', False)
        
        # Set initial sash position to ensure visibility (30% for top frame, 70% for bottom)
        self.daily_usage_paned.after(200, self._ensure_daily_usage_sash_visibility)
        self.daily_usage_paned.after(500, self._ensure_daily_usage_sash_visibility)
        self.daily_usage_paned.after(1000, self._ensure_daily_usage_sash_visibility)
        self.daily_usage_paned.after(1200, self._ensure_canvas_scroll_region)

        
        entry_frame = ttk.LabelFrame(self.daily_usage_paned, text="현장별 일일 사용량 기입")
        self.daily_usage_paned.add(entry_frame, weight=1) # Changed from weight=3 to weight=1
        
        # Header area with two rows to prevent buttons from being hidden on small screens
        header_container = ttk.Frame(entry_frame)
        header_container.pack(fill='x', padx=2, pady=1)
        
        row1 = ttk.Frame(header_container)
        row1.pack(fill='x', pady=1)
        row2 = ttk.Frame(header_container)
        row2.pack(fill='x', pady=1)
        
        # Row 1: Primary Actions
        self.btn_daily_save = ttk.Button(row1, text="💾 저장", style='Action.TButton', command=self.add_daily_usage_entry, width=8)
        self.btn_daily_save.pack(side='left', padx=2)
        
        ttk.Button(row1, text="🧹 초기화", command=self.clear_daily_usage_form_all, width=10).pack(side='left', padx=2)
        
        btn_ndt_map = ttk.Button(row1, text="🧪 NDT 품목 매핑", command=self.open_ndt_product_map_dialog)
        btn_ndt_map.pack(side='left', padx=5)

        btn_sync = ttk.Button(row1, text="🔄 작업자 일괄 적용", command=self.sync_worker_times, width=20)
        btn_sync.pack(side='left', padx=5)

        self.btn_daily_report = ttk.Button(row1, text="📄 작업일보 출력", command=self.export_daily_work_report, width=15)
        self.btn_daily_report.pack(side='left', padx=5)

        btn_report_map = ttk.Button(row1, text="⚙️ 출력 설정", command=self.open_report_mapping_dialog)
        btn_report_map.pack(side='left', padx=5)

        self.btn_sash_lock = ttk.Button(row1, text="🔒 경계 고정됨" if self.daily_usage_sash_locked else "🔓 경계 고정", command=self.toggle_sash_lock)
        self.btn_sash_lock.pack(side='right', padx=5)

        # Row 2: Secondary / Tool Actions
        btn_save_sess = ttk.Button(row2, text="💾 세션 저장", command=self.save_form_session, width=12)
        btn_save_sess.pack(side='left', padx=5)
        
        btn_load_sess = ttk.Button(row2, text="📂 세션 불러오기", command=self.load_form_session, width=15)
        btn_load_sess.pack(side='left', padx=5)
        
        btn_add_vehicle = ttk.Button(row2, text="🚗 차량점검", command=self.add_vehicle_inspection_box)
        btn_add_vehicle.pack(side='right', padx=2)

        btn_add_checklist = ttk.Button(row2, text="☑️ 목록 추가", command=self.add_new_checklist)
        btn_add_checklist.pack(side='right', padx=2)

        btn_add_memo = ttk.Button(row2, text="➕ 메모 추가", command=self.add_new_memo)
        btn_add_memo.pack(side='right', padx=2)
        if self.daily_usage_sash_locked:
            try:
                self.style.configure("SashLock.TButton", foreground="red")
                self.btn_sash_lock.configure(style="SashLock.TButton")
            except Exception: pass

        # [STABILITY FIX] Use a Canvas to hold the entry form for draggable support.
        # [REVISION] Restored scrollbar to satisfy the user request.
        canvas_parent = ttk.Frame(entry_frame)
        canvas_parent.pack(fill='both', expand=True, padx=2, pady=1)
        
        self.entry_canvas = tk.Canvas(canvas_parent, highlightthickness=0, bg=self.theme_bg)
        entry_vsb = ttk.Scrollbar(canvas_parent, orient="vertical", command=self.entry_canvas.yview)
        entry_vsb.pack(side='right', fill='y')
        self.entry_canvas.configure(yscrollcommand=entry_vsb.set)
        self.entry_canvas.pack(side='left', fill='both', expand=True)
        
        self.entry_inner_frame = ttk.Frame(self.entry_canvas)
        # Use a window to place the frame inside the canvas
        self.entry_canvas_window = self.entry_canvas.create_window((0, 0), window=self.entry_inner_frame, anchor='nw')
        
        def _on_entry_config(e):
            # Update canvas window width: use canvas width but ensure a minimum of 1100 
            # to fit both form (550) and workers (550+) without clipping.
            target_w = max(1100, e.width)
            self.entry_canvas.itemconfig(self.entry_canvas_window, width=target_w)
            self._ensure_canvas_scroll_region()
        
        self.entry_inner_frame.bind("<Configure>", lambda e: self._ensure_canvas_scroll_region())
        self.entry_canvas.bind("<Configure>", _on_entry_config)
        
        # [NOTE] Mousewheel scrolling is handled by the global handler in __init__

        
        # Explicitly fix all possible grid rows to weight 0
        for r in range(100):
            self.entry_inner_frame.grid_rowconfigure(r, weight=0)
        self.entry_inner_frame.columnconfigure(0, weight=1)
        
        # 1. Unified Master Form Panel summerly
        self.master_form_panel = ttk.LabelFrame(self.entry_inner_frame, text="일일 검사 및 사용량 기록")
        self.master_form_panel.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)

        
        # Configure columns inside the master panel (Reduced minsize)
        self.master_form_panel.columnconfigure(0, weight=0, minsize=350)
        self.master_form_panel.columnconfigure(1, weight=1)
        
                # Inner content for the basic form
        form_content = ttk.Frame(self.master_form_panel, padding=10)
        form_content.grid(row=0, column=0, sticky='w')
        
        for c in range(4): form_content.columnconfigure(c, weight=0)

        # Row 0: 업체명, 현장명
        ttk.Label(form_content, text="업체명:").grid(row=0, column=0, padx=(5, 0), pady=1, sticky='e')
        co_container = ttk.Frame(form_content)
        co_container.grid(row=0, column=1, padx=(2, 10), pady=1, sticky='w')
        self.cb_daily_company = ttk.Combobox(co_container, width=12, values=self.companies)
        self.cb_daily_company.pack(side='left')
        btn_company_mgr = tk.Button(co_container, text="⚙️ 관리", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                                   command=lambda: self.open_list_management_dialog('companies', target_cb=self.cb_daily_company))
        btn_company_mgr.pack(side='left', padx=(2, 0))

        ttk.Label(form_content, text="현장명:").grid(row=0, column=2, padx=(5, 0), pady=1, sticky='e')
        site_container = ttk.Frame(form_content)
        site_container.grid(row=0, column=3, padx=(2, 5), pady=1, sticky='w')
        self.cb_daily_site = ttk.Combobox(site_container, width=12, values=self.sites)
        self.cb_daily_site.pack(side='left')
        btn_site_mgr = tk.Button(site_container, text="⚙️ 관리", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                                command=lambda: self.open_list_management_dialog('sites', target_cb=self.cb_daily_site))
        btn_site_mgr.pack(side='left', padx=(2, 0))

        # Row 1: 날짜, 장비명
        ttk.Label(form_content, text="날짜:").grid(row=1, column=0, padx=(5, 0), pady=1, sticky='e')
        from tkcalendar import DateEntry
        self.ent_daily_date = DateEntry(form_content, width=15, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly')
        self.ent_daily_date.grid(row=1, column=1, padx=(2, 10), pady=1, sticky='w')

        ttk.Label(form_content, text="장비명:").grid(row=1, column=2, padx=(5, 0), pady=1, sticky='e')
        equip_container = ttk.Frame(form_content)
        equip_container.grid(row=1, column=3, padx=(2, 5), pady=1, sticky='w')
        
        self.ent_daily_equip = ttk.Entry(equip_container, width=15)
        self.ent_daily_equip.pack(side='left', fill='x', expand=True)
        
        # Link to the old variable name to avoid breaking other parts of the code
        self.cb_daily_equip = self.ent_daily_equip 
        
        # [MODERN] Place the search button INSIDE the entry widget at the right end
        btn_equip_search = tk.Button(self.ent_daily_equip, text="🔍", font=('Arial', 8), 
                                    bd=0, bg='white', cursor='hand2',
                                    command=self.open_equipment_search_dialog)
        btn_equip_search.place(relx=1.0, x=-2, rely=0.5, anchor='e', width=18, height=18)
        
        btn_equip_mgr = tk.Button(equip_container, text="⚙️ 관리", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                                 command=lambda: self.open_list_management_dialog('equipments', target_cb=self.ent_daily_equip))
        btn_equip_mgr.pack(side='left', padx=(2, 0))

        # Row 2: 품목명
        ttk.Label(form_content, text="품목명:").grid(row=2, column=0, padx=(5, 0), pady=1, sticky='e')
        mat_container = ttk.Frame(form_content)
        mat_container.grid(row=2, column=1, columnspan=3, padx=(2, 5), pady=1, sticky='w')
        
        self.cb_daily_material = ttk.Entry(mat_container, width=35)
        self.cb_daily_material.pack(side='left')
        
        # [MODERN] Place the search button INSIDE the entry
        btn_material_search = tk.Button(self.cb_daily_material, text="🔍", font=('Arial', 8), 
                                       bd=0, bg='white', cursor='hand2',
                                       command=lambda: self.open_material_search_dialog(target_form='daily_usage'))
        btn_material_search.place(relx=1.0, x=-2, rely=0.5, anchor='e', width=18, height=18)
        
        btn_material_mgr = tk.Button(mat_container, text="⚙️ 관리", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                                    command=lambda: self.open_list_management_dialog('materials', target_cb=self.cb_daily_material))
        btn_material_mgr.pack(side='left', padx=(2, 0))

        # Row 3: 방법, 검사품명
        ttk.Label(form_content, text="방법:").grid(row=3, column=0, padx=(5, 0), pady=1, sticky='e')
        self.cb_daily_test_method = ttk.Combobox(form_content, width=15, values=['RT', 'PAUT', 'UT', 'MT', 'PT', 'ETC'])
        self.cb_daily_test_method.grid(row=3, column=1, padx=(2, 10), pady=1, sticky='w')
        
        ttk.Label(form_content, text="검사품명:").grid(row=3, column=2, padx=(5, 0), pady=1, sticky='e')
        insp_container = ttk.Frame(form_content)
        insp_container.grid(row=3, column=3, padx=(2, 5), pady=1, sticky='w')
        self.ent_daily_inspection_item = ttk.Entry(insp_container, width=15)
        self.ent_daily_inspection_item.pack(side='left')
        
        btn_insp_item_mgr = tk.Button(insp_container, text="⚙️ 관리", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                                     command=lambda: self.open_list_management_dialog('test_items', target_cb=self.ent_daily_inspection_item))
        btn_insp_item_mgr.pack(side='left', padx=(2, 0))

        # Row 4: 수량, 단위
        ttk.Label(form_content, text="수량:").grid(row=4, column=0, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_test_amount = ttk.Entry(form_content, width=15)
        self.ent_daily_test_amount.grid(row=4, column=1, padx=(2, 10), pady=1, sticky='w')
        
        ttk.Label(form_content, text="단위:").grid(row=4, column=2, padx=(5, 0), pady=1, sticky='e')
        unit_container = ttk.Frame(form_content)
        unit_container.grid(row=4, column=3, padx=(2, 5), pady=1, sticky='w')
        self.cb_daily_unit = ttk.Combobox(unit_container, width=12, values=self.daily_units)
        self.cb_daily_unit.pack(side='left')
        
        btn_unit_mgr = tk.Button(unit_container, text="⚙️ 관리", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                                command=lambda: self.open_list_management_dialog('daily_units', target_cb=self.cb_daily_unit))
        btn_unit_mgr.pack(side='left', padx=(2, 0))

        # Row 5: 단가, 출장비
        ttk.Label(form_content, text="단가:").grid(row=5, column=0, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_unit_price = ttk.Entry(form_content, width=15)
        self.ent_daily_unit_price.grid(row=5, column=1, padx=(2, 10), pady=1, sticky='w')

        ttk.Label(form_content, text="출장비:").grid(row=5, column=2, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_travel_cost = ttk.Entry(form_content, width=15)
        self.ent_daily_travel_cost.grid(row=5, column=3, padx=(2, 5), pady=1, sticky='w')

        # Row 6: 적용코드, 성적서번호
        ttk.Label(form_content, text="적용코드:").grid(row=6, column=0, padx=(5, 0), pady=1, sticky='e')
        app_container = ttk.Frame(form_content)
        app_container.grid(row=6, column=1, padx=(2, 10), pady=1, sticky='w')
        self.ent_daily_applied_code = ttk.Entry(app_container, width=12)
        self.ent_daily_applied_code.pack(side='left')
        
        btn_app_code_mgr = tk.Button(app_container, text="⚙️ 관리", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                                    command=lambda: self.open_list_management_dialog('applied_codes', target_cb=self.ent_daily_applied_code))
        btn_app_code_mgr.pack(side='left', padx=(2, 0))

        ttk.Label(form_content, text="성적서번호:").grid(row=6, column=2, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_report_no = ttk.Entry(form_content, width=18)
        self.ent_daily_report_no.grid(row=6, column=3, padx=(2, 5), pady=1, sticky='w')

        # Row 7: 비고, 일식
        ttk.Label(form_content, text="비고:").grid(row=7, column=0, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_note = ttk.Entry(form_content, width=15)
        self.ent_daily_note.grid(row=7, column=1, padx=(2, 10), pady=1, sticky='w')

        ttk.Label(form_content, text="일식:").grid(row=7, column=2, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_meal_cost = ttk.Entry(form_content, width=15)
        self.ent_daily_meal_cost.grid(row=7, column=3, padx=(2, 5), pady=1, sticky='w')

        # Row 8: 검사비
        ttk.Label(form_content, text="검사비:").grid(row=8, column=0, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_test_fee = ttk.Entry(form_content, width=15)
        self.ent_daily_test_fee.grid(row=8, column=1, padx=(2, 10), pady=1, sticky='w')

        
        # Restore focus transitions and defaults
        self.ent_daily_inspection_item.insert(0, "Piping")
        self.ent_daily_inspection_item.bind('<Return>', lambda e: self.ent_daily_test_amount.focus_set())
        self.ent_daily_test_amount.bind('<Return>', lambda e: self.cb_daily_unit.focus_set())
        self.cb_daily_unit.set('매')
        self.cb_daily_unit.bind('<Return>', lambda e: self.ent_daily_unit_price.focus_set())
        self.cb_daily_unit.bind('<<ComboboxSelected>>', lambda e: self.ent_daily_unit_price.focus_set())
        self.ent_daily_unit_price.bind('<Return>', lambda e: self.ent_daily_applied_code.focus_set())
        self.ent_daily_applied_code.insert(0, "KS")
        self.ent_daily_applied_code.bind('<Return>', lambda e: self.ent_daily_report_no.focus_set())
        self.ent_daily_report_no.bind('<Return>', lambda e: self.ent_daily_note.focus_set())
        self.ent_daily_note.bind('<Return>', lambda e: self.ent_daily_meal_cost.focus_set())
        self.ent_daily_meal_cost.insert(0, "0")
        self.ent_daily_meal_cost.bind('<Return>', lambda e: self.ent_daily_test_fee.focus_set())
        
        def on_method_select_focus(e):
            self.root.after(10, self.ent_daily_inspection_item.focus_set)
            return "break"
        self.cb_daily_test_method.bind('<<ComboboxSelected>>', on_method_select_focus, add='+')
        self.cb_daily_test_method.bind('<Return>', on_method_select_focus, add='+')
        
        def on_method_change_auto_unit_logic(e):
            method = self.cb_daily_test_method.get().strip()
            unit_map = {'RT': '매', 'UT': 'P,M,I/D', 'MT': 'P,M,I/D', 'PT': 'P,M,I/D', 'PAUT': 'M,I/D'}
            if method in unit_map: self.cb_daily_unit.set(unit_map[method])
            return "break"
        self.cb_daily_test_method.bind('<<ComboboxSelected>>', on_method_change_auto_unit_logic, add='+')
        
        # [ROBUST_AUTOCOMPLETE] Use standardized suggest system for Entry Form
        self._bind_combobox_word_suggest(self.cb_daily_site, lambda: sorted(list(set(self.sites))))
        self._bind_combobox_word_suggest(self.cb_daily_material, lambda: self._get_material_candidates(include_all=False))
        self._bind_combobox_word_suggest(self.cb_daily_equip, lambda: self._get_equipment_candidates(include_all=False))
        self._bind_combobox_word_suggest(self.ent_daily_inspection_item, lambda: self._get_inspection_item_candidates())
        self._bind_combobox_word_suggest(self.ent_daily_applied_code, lambda: self._get_applied_code_candidates())

        # Category definitions for focus flow and loops
        ndt_materials = self.ndt_materials_all
        rtk_cats = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타", "총계"]



        # Row 1: NDT with Multi-Company Support
        self.ndt_frame = ttk.LabelFrame(self.master_form_panel, text="NDT 자재 소모량 (회사별)")
        self.ndt_frame.grid(row=1, column=0, padx=5, pady=2, sticky='ew')
        
        # Container for company-specific NDT sections
        self.ndt_company_container = ttk.Frame(self.ndt_frame)
        self.ndt_company_container.pack(fill='x', expand=True, padx=5, pady=2)
        
        # Store NDT entries by company index: {0: {mat: entry, ...}, 1: {...}}
        self.ndt_company_entries = []
        
        # Button frame for adding companies
        btn_frame = ttk.Frame(self.ndt_frame)
        btn_frame.pack(fill='x', padx=5, pady=2)
        ttk.Button(btn_frame, text="+ 회사 추가", command=self.add_ndt_company_section, width=12).pack(side='left', padx=2)
        ttk.Button(btn_frame, text="- 마지막 회사 삭제", command=self.remove_last_ndt_company, width=15).pack(side='left', padx=2)
        
        # Add default first company section
        self.add_ndt_company_section()

        # Removed draggable container logic

        # Row 2: RTK
        self.rtk_grid = ttk.LabelFrame(self.master_form_panel, text="RTK 분류")
        self.rtk_grid.grid(row=2, column=0, padx=5, pady=2, sticky='ew')
        
        for c in range(6): self.rtk_grid.columnconfigure(c, weight=1, uniform="ndt_rtk")
        self.rtk_entries = {}
        for i, cat in enumerate(rtk_cats):
            r = i // 3; col = (i % 3) * 2
            ttk.Label(self.rtk_grid, text=f"{cat}:", font=('Arial', 8)).grid(row=r, column=col, padx=1, pady=1, sticky='w')
            e = ttk.Entry(self.rtk_grid, width=6)
            e.grid(row=r, column=col+1, padx=1, pady=1, sticky='ew')
            self.rtk_entries[cat] = e
            
            # Focus transition to next RTK entry
            if i + 1 < len(rtk_cats) - 1: # Skip "총계" (the last one)
                next_cat = rtk_cats[i+1]
                e.bind('<Return>', lambda e, nc=next_cat: self.rtk_entries[nc].focus_set())
            elif cat != "총계": # From last editable RTK to save button container (approx)
                # We don't have a direct handle to the button, but we can focus the first worker's name
                e.bind('<Return>', lambda e: self.cb_daily_user.focus_set())
            
            # Bind auto-calculation
            if cat != "총계":
                e.bind('<KeyRelease>', lambda e: self.calculate_rtk_total())
        
        self.rtk_entries["총계"].config(state='readonly')
        # Removed draggable container logic

        # [LAYOUT FIX] Ensure enough baseline row heights
        try:
            self.entry_inner_frame.grid_rowconfigure(0, minsize=170)
            self.entry_inner_frame.grid_rowconfigure(1, minsize=120)
            self.entry_inner_frame.grid_rowconfigure(2, minsize=130)
            # Row 3 (Extra spacer) removed to minimize whitespace
        except Exception:
            pass


        # [MIGRATION] Convert existing times to "익일" format and sort
        self.worktimes = self._migrate_worktimes(self.worktimes if hasattr(self, 'worktimes') else [])
        
        # Create a single container for all workers inside the master panel summerly
        workers_box_frame = ttk.Frame(self.master_form_panel)
        workers_box_frame.grid(row=0, column=1, rowspan=3, sticky='nsew', padx=5, pady=5)
        workers_box_frame.columnconfigure(0, weight=1)
        workers_box_frame.rowconfigure(1, weight=1)
        
        worker_hdr = ttk.Frame(workers_box_frame)
        worker_hdr.grid(row=0, column=0, sticky='ew', pady=(0, 2))
        ttk.Label(worker_hdr, text="작업자 기록", font=('Malgun Gothic', 9, 'bold')).pack(side='left')
        ttk.Button(worker_hdr, text="⚙", width=2, 
                   command=lambda: self.open_list_management_dialog('users')).pack(side='right')

        workers_inner = ttk.Frame(workers_box_frame)
        workers_inner.grid(row=1, column=0, sticky='nsew')
        # Configure inner grid for workers (2 columns)
        for c in range(2): workers_inner.grid_columnconfigure(c, weight=1)
        
        def setup_worker_group(idx, row, col):
            # Each worker group is no longer individually draggable
            # Instead, they are sub-frames within the main workers_inner
            group_frame = ttk.LabelFrame(workers_inner, text=f"작업자 {idx}")
            group_frame.grid(row=row, column=col, padx=2, pady=1, sticky='ew')
            
            group = WorkerDataGroup(
                group_frame, worker_index=idx, users_list=getattr(self, 'users', []),
                enable_autocomplete=True,
                time_list=self.worktimes
            )
            group.pack(fill='x', expand=True, padx=1, pady=1)

            # Store global references for all workers (1-10)
            if idx == 1:
                self.cb_daily_user = group.composite
                self.ent_worktime1 = group.ent_worktime
                self.ent_ot1 = group.ent_ot
            
            # Always set index-specific attributes for sync/data access
            setattr(self, f'worker_group{idx}', group)
            setattr(self, f'cb_daily_user{idx}', group.composite)
            setattr(self, f'ent_worktime{idx}', group.ent_worktime)
            setattr(self, f'ent_ot{idx}', group.ent_ot)
            
            # Bindings for auto-save & auto-OT
            group.bind_name('<FocusOut>', lambda e: self.auto_save_to_list(e, group.cb_name, self.users, 'users'))
            group.bind_name('<Return>', lambda e: self.auto_save_to_list(e, group.cb_name, self.users, 'users'))
            
            group.bind_time('<FocusOut>', lambda e: self.auto_save_worktime(e, group.ent_worktime, 'worktimes'))
            group.bind_time('<Return>', lambda e: self.auto_save_worktime(e, group.ent_worktime, 'worktimes'))
            
            group.bind_ot('<FocusOut>', lambda e: self.auto_save_ot(e, group.ent_ot, 'ot_times'))
            group.bind_ot('<Return>', lambda e: self.auto_save_ot(e, group.ent_ot, 'ot_times'))
            
            return group_frame, group

        # 5 rows x 2 cols = 10 workers rearranged inside workers_inner
        for i in range(1, 6): setup_worker_group(i, i-1, 0)
        for i in range(6, 11): setup_worker_group(i, i-6, 1)

        # Removed draggable worker box logic

        # Bindings & Finalization
        calc_trigger = lambda e: self.update_daily_test_fee_calc()
        
        def on_qty_change(e):
            self.update_daily_test_fee_calc()
            
        # Handle Date Changes globally for recalculations
        def on_date_change(e=None):
            # Recalculate OT for all workers when date changes (weekend vs weekday rates)
            for i in range(1, 11):
                group = getattr(self, f'worker_group{i}', None)
                if group:
                    self.calculate_and_update_ot(group.ent_worktime.get(), group.ent_worktime)

        self.ent_daily_date.bind('<<DateEntrySelected>>', on_date_change, add='+')
        
        # Initial call to set last known date for change detection
        try:
            self._last_daily_date = self.ent_daily_date.get_date()
        except:
            self._last_daily_date = None
        
        # [V19.2_AUTO_LOAD] Populate the list automatically on startup/setup
        self.root.after(1500, self.update_daily_usage_view)
        
        self.ent_daily_test_amount.bind('<KeyRelease>', on_qty_change)
        
        # [NEW] Add comma auto-formatting and Focus Transition on FocusOut/Return
        cost_entries = [
            self.ent_daily_test_amount,
            self.ent_daily_unit_price, 
            self.ent_daily_travel_cost, 
            self.ent_daily_test_fee
        ]
        
        for i, ent in enumerate(cost_entries):
            ent.bind('<KeyRelease>', calc_trigger, add='+')
            ent.bind('<FocusOut>', lambda e, widget=ent: self.format_entry_with_commas(e, widget), add='+')
            
            # Define focus transition
            def on_return(e, current_idx=i):
                # Format current entry first
                self.format_entry_with_commas(e, cost_entries[current_idx])
                # Move focus to next entry if possible
                if current_idx + 1 < len(cost_entries):
                    cost_entries[current_idx + 1].focus_set()
                return "break"
                
            ent.bind('<Return>', on_return)
        
        self.update_material_combo()
        self._adjust_parent_height(self.entry_inner_frame, force=True)
        
        
        display_frame = ttk.LabelFrame(self.daily_usage_paned, text="일일 사용량 기록 조회")
        self.daily_usage_paned.add(display_frame, weight=1) # Less weight for the list
        
        # Filter controls
        filter_frame = ttk.Frame(display_frame)
        filter_frame.pack(fill='x', padx=5, pady=5)
        
        # --- Row 1: Date Filters ---
        date_row = ttk.Frame(filter_frame)
        date_row.pack(fill='x', pady=2)
        
        ttk.Label(date_row, text="시작일:").pack(side='left', padx=5)
        self.ent_daily_start_date = DateEntry(date_row, width=12, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        self.ent_daily_start_date.pack(side='left', padx=5)
        # Default to showing all history (starting from 2024)
        start_date = datetime.datetime(2024, 1, 1)
        self.ent_daily_start_date.set_date(start_date)
        
        ttk.Label(date_row, text="종료일:").pack(side='left', padx=5)
        self.ent_daily_end_date = DateEntry(date_row, width=12, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        self.ent_daily_end_date.pack(side='left', padx=5)
        self.ent_daily_end_date.set_date(datetime.datetime.now())
        
        # --- Row 2: Search Filters ---
        filter_row = ttk.Frame(filter_frame)
        filter_row.pack(fill='x', pady=2)
        
        ttk.Label(filter_row, text="업체명:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_company = ttk.Combobox(filter_row, width=12)
        self.cb_daily_filter_company.pack(side='left', padx=2)
        self.cb_daily_filter_company.set('전체')
        tk.Button(filter_row, text="⚙️", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                  command=lambda: self.open_list_management_dialog('companies', target_cb=self.cb_daily_filter_company)).pack(side='left', padx=(0, 5))
        
        ttk.Label(filter_row, text="현장:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_site = ttk.Combobox(filter_row, width=12)
        self.cb_daily_filter_site.pack(side='left', padx=2)
        self.cb_daily_filter_site.set('전체')
        tk.Button(filter_row, text="⚙️", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                  command=lambda: self.open_list_management_dialog('sites', target_cb=self.cb_daily_filter_site)).pack(side='left', padx=(0, 5))
        self._bind_combobox_word_suggest(self.cb_daily_filter_site, lambda: ['전체'] + sorted(list(set(self.sites))))
        
        ttk.Label(filter_row, text="품목명:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_material = ttk.Combobox(filter_row, width=15)
        self.cb_daily_filter_material.pack(side='left', padx=2)
        self.cb_daily_filter_material.set('전체')
        tk.Button(filter_row, text="⚙️", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                  command=lambda: self.open_list_management_dialog('materials', target_cb=self.cb_daily_filter_material)).pack(side='left', padx=(0, 5))
        self._bind_combobox_word_suggest(self.cb_daily_filter_material, lambda: self._get_material_candidates(include_all=True))
        
        ttk.Label(filter_row, text="장비명:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_equipment = ttk.Combobox(filter_row, width=15)
        self.cb_daily_filter_equipment.pack(side='left', padx=2)
        self.cb_daily_filter_equipment.set('전체')
        tk.Button(filter_row, text="⚙️", font=('Malgun Gothic', 8), bd=0, bg=self.theme_bg, fg='blue', cursor='hand2',
                  command=lambda: self.open_list_management_dialog('equipments', target_cb=self.cb_daily_filter_equipment)).pack(side='left', padx=(0, 5))
        self._bind_combobox_word_suggest(self.cb_daily_filter_equipment, lambda: self._get_equipment_candidates(include_all=True))
        
        ttk.Label(filter_row, text="작업자:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_worker = ttk.Combobox(filter_row, width=10)
        self.cb_daily_filter_worker.pack(side='left', padx=2)
        self.cb_daily_filter_worker.set('전체')
        
        ttk.Label(filter_row, text="분류:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_shift = ttk.Combobox(filter_row, width=8, state="readonly", values=["전체", "주간", "야간", "주야간", "휴일"])
        self.cb_daily_filter_shift.pack(side='left', padx=2)
        self.cb_daily_filter_shift.set('전체')
        
        # --- Row 3: Action Buttons ---
        btn_row = ttk.Frame(filter_frame)
        btn_row.pack(fill='x', pady=5)
        
        btn_filter = ttk.Button(btn_row, text="조회", style='Action.TButton', command=self.update_daily_usage_view)
        btn_filter.pack(side='left', padx=5)
        
        btn_filter_reset = ttk.Button(btn_row, text="♻️ 필터 초기화", command=self.reset_daily_usage_filters)
        btn_filter_reset.pack(side='left', padx=5)
        
        btn_delete = ttk.Button(btn_row, text="선택 항목 삭제", command=self.delete_daily_usage_entry)
        btn_delete.pack(side='left', padx=10)
        
        btn_edit = ttk.Button(btn_row, text="선택 항목 수정", command=self.open_edit_daily_usage_dialog)
        btn_edit.pack(side='left', padx=5)
        
        btn_export = ttk.Button(btn_row, text="엑셀 내보내기", command=self.export_daily_usage_history)
        btn_export.pack(side='left', padx=5)
        
        btn_export_all = ttk.Button(btn_row, text="전체 기록 내보내기", command=self.export_all_daily_usage)
        btn_export_all.pack(side='left', padx=5)
        
        btn_col_manage = ttk.Button(btn_row, text="컬럼 관리", command=self.show_column_visibility_dialog)
        btn_col_manage.pack(side='left', padx=10)

        # Dedicated Save Button for the List View
        self.btn_daily_save_list = ttk.Button(btn_row, text="💾 변경사항 저장", command=self.save_all_daily_usage_changes, style='Accent.TButton' if 'Accent.TButton' in self.style.theme_names() else 'TButton')
        self.btn_daily_save_list.pack(side='left', padx=10)

        # Bindings
        filter_widgets = [
            self.cb_daily_filter_site, self.cb_daily_filter_company, self.cb_daily_filter_material, 
            self.cb_daily_filter_equipment, self.cb_daily_filter_worker, 
            self.cb_daily_filter_shift
        ]
        for widget in filter_widgets:
            widget.bind("<Return>", lambda e: self.update_daily_usage_view())
            widget.bind("<<ComboboxSelected>>", lambda e: self.update_daily_usage_view())

        for date_widget in [self.ent_daily_start_date, self.ent_daily_end_date]:
            date_widget.bind("<<DateEntrySelected>>", lambda e: self.update_daily_usage_view())
            try:
                date_widget.bind("<Return>", lambda e: self.update_daily_usage_view())
                for child in date_widget.winfo_children():
                    if isinstance(child, (tk.Entry, ttk.Entry)):
                        child.bind("<Return>", lambda e: self.update_daily_usage_view())
            except: pass
        
        # Treeview for daily usage records
        tree_container = ttk.Frame(display_frame)
        tree_container.pack(expand=True, fill='both', padx=5, pady=5)
        
        list_frame = ttk.Frame(tree_container)
        list_frame.pack(fill='both', expand=True)

        # Scrollbars
        vsb = ttk.Scrollbar(list_frame, orient="vertical")
        hsb = ttk.Scrollbar(list_frame, orient="horizontal")
        
        # Treeview with RTK categories and NDT materials
        # Note: Workers 1-10 columns are kept in the 'columns' tuple for data storage,
        # but we will only show a consolidated '작업자' in 'displaycolumns'.
        # Added '(Full작업자)' for Excel export backup.
        columns = ('날짜', '업체명', '적용코드', '현장', '검사품명', '성적서번호', '작업자', '작업시간', 'OT1', 'OT2', 'OT3', 'OT4', 'OT5', 'OT6', 'OT7', 'OT8', 'OT9', 'OT10', '장비명', '검사방법', '회사코드', '수량', '단위', '단가', '출장비', '일식', '검사비', 'OT시간', 'OT금액', '품목명', '센터미스', '농도', '마킹미스', '필름마크', '취급부주의', '고객불만', '기타', 'RTK총계', '형광자분', '흑색자분', '백색페인트', '침투제', '세척제', '현상제', '형광침투제', '비고', '입력시간', '차량번호', '주행거리', '차량점검', '차량비고', '(Full작업자)')
        self.daily_usage_tree = ttk.Treeview(list_frame, columns=columns, show='headings',
                                              yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Define display columns - hidden (Full작업자) and individual OT columns by default
        # [NEW] Also hide 'OT시간' and '필름매수' by default per user request
        visible_defaults = ['날짜', '업체명', '적용코드', '현장', '검사품명', '성적서번호', '작업자', '작업시간', '장비명', '검사방법', '수량', '단위', '단가', '출장비', '일식', '검사비', 'OT금액', '품목명', 'RTK총계', '비고', '입력시간']
        self.daily_usage_tree['displaycolumns'] = visible_defaults
        
        vsb.config(command=self.daily_usage_tree.yview)
        hsb.config(command=self.daily_usage_tree.xview)
        
        # Column configuration
        col_widths = {
            '날짜': 160, '현장': 130, '작업자': 170, '작업시간': 120,
            'OT1': 140, 'OT2': 140, 'OT3': 140, 'OT4': 140, 'OT5': 140, 'OT6': 140,
            'OT7': 140, 'OT8': 140, 'OT9': 140, 'OT10': 140,
            '차량번호': 120, '주행거리': 100, '차량점검': 200, '차량비고': 150,
            '장비명': 130, '검사방법': 90, '회사코드': 80, '적용코드': 100, '검사품명': 130, '성적서번호': 120,
            '수량': 70, '단위': 60, '단가': 90, '출장비': 90, '업체명': 120, '일식': 80, 
            '검사비': 100, 'OT시간': 80, 'OT금액': 100, '품목명': 210, '센터미스': 70, '농도': 70, '마킹미스': 70, 
            '필름마크': 70, '취급부주의': 70, '고객불만': 70, '기타': 70, 'RTK총계': 80, 
            '형광자분': 80, '흑색자분': 80, '백색페인트': 80, '침투제': 80, '세척제': 80, 
            '현상제': 80, '형광침투제': 80, '비고': 230, '입력시간': 300
        }
        
        for col in columns:
            self.daily_usage_tree.heading(col, text=col, command=lambda c=col: self.treeview_sort_column(self.daily_usage_tree, c, False))
            width = col_widths.get(col, 100)
            self.daily_usage_tree.column(col, width=width, minwidth=20, stretch=False, anchor='center')
        
        # Grid layout for list_frame
        self.daily_usage_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)
        
        # Auto-save column widths when user resizes columns
        # [NEW] Bind worker details popup AND Quick Edit dialog to double-click
        self.daily_usage_tree.bind("<Button-1>", lambda e: self.show_worker_popup(e, self.daily_usage_tree), add="+")
        self.daily_usage_tree.bind("<Double-1>", lambda e: self.show_worker_popup(e, self.daily_usage_tree), add="+")
        self.daily_usage_tree.bind("<Double-1>", lambda e: self.open_edit_daily_usage_dialog(), add="+")
        
        # [NEW] Enable column reordering via drag & drop
        self.enable_tree_column_drag(self.daily_usage_tree)
        
        def save_column_widths(event=None):
            self.save_tab_config()
            
        self.daily_usage_tree.bind('<ButtonRelease-1>', save_column_widths)

        # [NEW] Note Detail Area for long descriptions
        detail_frame = ttk.LabelFrame(display_frame, text="상세 비고 (선택된 항목)")
        detail_frame.pack(fill='x', padx=5, pady=(0, 5))
        
        self.txt_daily_note_detail = tk.Text(detail_frame, height=3, font=('Malgun Gothic', 10), wrap='word', bg='#F9F9F9')
        self.txt_daily_note_detail.pack(fill='both', expand=True, padx=5, pady=5)
        self.txt_daily_note_detail.config(state='disabled')

        # Selection binding to update detail area
        self.daily_usage_tree.bind('<<TreeviewSelect>>', self._on_daily_usage_select)
        
        # Optimize treeview scroll region to prevent unnecessary scrollbars
        def optimize_treeview_scroll_region():
            try:
                if hasattr(self, 'daily_usage_tree'):
                    self.daily_usage_tree.update_idletasks()
                    
                    # Get treeview content dimensions
                    total_items = len(self.daily_usage_tree.get_children())
                    if total_items > 0:
                        # Calculate approximate content height
                        item_height = 25  # Approximate height per row
                        content_height = total_items * item_height + 50  # Add some padding
                        
                        # Get treeview dimensions
                        tree_width = self.daily_usage_tree.winfo_width()
                        tree_height = self.daily_usage_tree.winfo_height()
                        
                        # Set scroll region to content size if smaller than treeview
                        if content_height < tree_height:
                            self.daily_usage_tree.configure(yscrollcommand=(0, 0, tree_width, content_height))
                            print(f"Treeview scroll region optimized: (0, 0, {tree_width}, {content_height})")
                        else:
                            # Use full treeview dimensions
                            self.daily_usage_tree.configure(yscrollcommand=(0, 0, tree_width, tree_height))
                            print(f"Treeview scroll region set to full size: (0, 0, {tree_width}, {tree_height})")
            except Exception as e:
                print(f"Error optimizing treeview scroll region: {e}")
        
        # Apply optimization after a short delay
        self.root.after(500, optimize_treeview_scroll_region)

    def _on_daily_usage_select(self, event):
        """[NEW] Update Note Detail Area and load record to form when a row is selected in Site tab"""
        if not hasattr(self, 'daily_usage_tree'): return
        
        selection = self.daily_usage_tree.selection()
        if not selection:
            if hasattr(self, 'txt_daily_note_detail'):
                self.txt_daily_note_detail.config(state='normal')
                self.txt_daily_note_detail.delete('1.0', tk.END)
                self.txt_daily_note_detail.config(state='disabled')
            return
            
        item = selection[0]
        tags = self.daily_usage_tree.item(item, 'tags')
        if tags and tags[0].isdigit():
            idx = int(tags[0])
            if idx in self.daily_usage_df.index:
                row_data = self.daily_usage_df.loc[idx].to_dict()
                self.load_daily_usage_to_form(row_data)

        # Update note detail area as before
        if hasattr(self, 'txt_daily_note_detail'):
            values = self.daily_usage_tree.item(item, 'values')
            if values:
                try:
                    cols = self.daily_usage_tree['columns']
                    if '비고' in cols:
                        note_idx = list(cols).index('비고')
                        if note_idx < len(values):
                            note_text = values[note_idx]
                            self.txt_daily_note_detail.config(state='normal')
                            self.txt_daily_note_detail.delete('1.0', tk.END)
                            self.txt_daily_note_detail.insert(tk.END, note_text)
                            self.txt_daily_note_detail.config(state='disabled')
                except Exception as e:
                    print(f"Detail view error: {e}")


    


    def export_budget_sales_status(self):
        """Export the currently filtered budget sales status to Excel"""
        if not hasattr(self, 'budget_view_tree') or not self.budget_view_tree.get_children():
            messagebox.showwarning("데이터 없음", "내보낼 데이터가 없습니다. 먼저 조회를 해주세요.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"매출현황_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not file_path:
            return

        # Collect data from treeview
        data = []
        columns = [self.budget_view_tree.heading(col)['text'] for col in self.budget_view_tree['columns']]
        for item in self.budget_view_tree.get_children():
            data.append(self.budget_view_tree.item(item)['values'])
            
        df = pd.DataFrame(data, columns=columns)
        
        # Add summary row at the bottom
        summary_row = {col: '' for col in columns}
        summary_row['날짜'] = '합계'
        if hasattr(self, 'lbl_bv_actual_income'):
            summary_row['검사단가'] = self.lbl_bv_actual_income.cget('text').replace(' 원', '').replace(',', '').strip()
        
        # Compute OT/Total directly from columns since Budget KPIs might show differently
        ot_idx = columns.index('OT합계') if 'OT합계' in columns else -1
        total_idx = columns.index('합계') if '합계' in columns else -1
        
        sum_ot = 0
        sum_tot = 0
        for _d in data:
            if ot_idx >= 0:
                try: sum_ot += float(str(_d[ot_idx]).replace(',', ''))
                except: pass
            if total_idx >= 0:
                try: sum_tot += float(str(_d[total_idx]).replace(',', ''))
                except: pass

        summary_row['OT합계'] = f"{sum_ot:g}"
        summary_row['합계'] = f"{sum_tot:g}"
        
        # Add the summary row to the data list for the DataFrame
        data.append([summary_row.get(col, '') for col in columns])
        
        df = pd.DataFrame(data, columns=columns)
        
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='SalesStatus', index=False)
                # Auto-adjust columns
                worksheet = writer.sheets['SalesStatus']
                for i, col in enumerate(df.columns):
                    max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                    worksheet.column_dimensions[chr(65 + i)].width = max_len
            
            messagebox.showinfo("내보내기 완료", "매출현황이 엑셀로 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 저장 중 오류가 발생했습니다: {e}")

    def setup_budget_tab(self):
        """Setup the project execution budget management tab with detailed labor cost breakdown"""
        # KPI Summary Panel [NEW]
        kpi_frame = tk.Frame(self.tab_budget, background="#ffffff", highlightthickness=1, highlightbackground="#e5e7eb")
        kpi_frame.pack(fill='x', padx=10, pady=(5, 0))
        
        kpis = [
            ("계약금액", "lbl_kpi_rev", "#374151"),
            ("실행원가", "lbl_kpi_cost", "#374151"),
            ("영업이익", "lbl_kpi_profit", "#10b981"),
            ("이익률", "lbl_kpi_margin", "#10b981")
        ]
        
        for i, (label, attr, color) in enumerate(kpis):
            f = tk.Frame(kpi_frame, background="#ffffff")
            f.pack(side='left', expand=True, fill='both', pady=10)
            if i > 0:
                tk.Frame(kpi_frame, width=1, background="#e5e7eb").pack(side='left', fill='y', pady=10)
            
            tk.Label(f, text=label, font=("Malgun Gothic", 9), background="#ffffff", foreground="#6b7280").pack()
            lbl_val = tk.Label(f, text="0원", font=("Malgun Gothic", 12, "bold"), background="#ffffff", foreground=color)
            lbl_val.pack()
            setattr(self, attr, lbl_val)

        # Fixed Button Frame [NEW]
        top_btn_frame = ttk.Frame(self.tab_budget)
        top_btn_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Button(top_btn_frame, text="실적 데이터로 채우기", command=self.fill_budget_from_actuals).pack(side='left', padx=5)
        ttk.Button(top_btn_frame, text="저장/수정", command=self.save_budget_entry).pack(side='left', padx=5)
        ttk.Button(top_btn_frame, text="삭제", command=self.delete_budget_entry).pack(side='left', padx=5)
        ttk.Button(top_btn_frame, text="🧹 초기화", command=self.clear_budget_form).pack(side='left', padx=5)
        
        # Pop-out button
        ttk.Button(top_btn_frame, text="🔍 팝업창으로 열기", command=self.open_detached_budget_view).pack(side='left', padx=15)
        
        ttk.Button(top_btn_frame, text="엑셀 내보내기", command=self.export_budget_sales_status).pack(side='right', padx=10)

        main_paned = ttk.PanedWindow(self.tab_budget, orient='vertical')
        main_paned.pack(fill='both', expand=True, padx=10, pady=(0, 5))
        
        # --- 상단: Scrollable Form area ---
        top_container = ttk.Frame(main_paned)
        main_paned.add(top_container, weight=2)
        
        canvas = tk.Canvas(top_container, highlightthickness=0)
        v_scroll = ttk.Scrollbar(top_container, orient="vertical", command=canvas.yview)
        form_scrollable = ttk.Frame(canvas)
        
        form_scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        _canvas_win = canvas.create_window((0, 0), window=form_scrollable, anchor="nw")
        canvas.configure(yscrollcommand=v_scroll.set)
        
        # [FIX] 우측 빈 공간 제거: Canvas 너비 변경 시 내부 프레임 너비를 동기화
        def _on_canvas_resize(event):
            canvas.itemconfigure(_canvas_win, width=event.width)
        canvas.bind("<Configure>", _on_canvas_resize)
        
        # Scroll binding is handled globally in MaterialManager.__init__
        
        canvas.pack(side="left", fill="both", expand=True)
        v_scroll.pack(side="right", fill="y")
        
        # 1. Main Budget Form
        form_frame = ttk.LabelFrame(form_scrollable, text="실행예산 입력/수정", padding=10)
        form_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        # Grid layout for form fields
        rows = [
            ("현장명:", "cb_budget_site"),
            ("계약금액 (Revenue):", "ent_budget_revenue"),
            ("매출금액 (UnitPrice):", "ent_budget_unit_price"),
            ("실행 노무비 (Labor):", "ent_budget_labor"),
            ("실행 재료비 (Material):", "ent_budget_material"),
            ("실행 경비 (Expense):", "ent_budget_expense"),
            ("실행 외주비 (Outsource):", "ent_budget_outsource"),
            ("영업이익 (Profit):", "ent_budget_profit"),
            ("이익률 (%):", "ent_budget_margin"),
            ("비고:", "ent_budget_note")
        ]
        
        self.budget_widgets = {}
        for i, (label_text, attr_name) in enumerate(rows):
            ttk.Label(form_frame, text=label_text).grid(row=i//4*2, column=i%4, sticky='w', padx=5, pady=2)
            if attr_name == "cb_budget_site":
                widget = ttk.Combobox(form_frame, width=30)
                widget['values'] = self.sites
            else:
                widget = ttk.Entry(form_frame, width=32)
                if attr_name == "ent_budget_labor":
                    # [STABILITY] Make labor cost read-only if it's auto-calculated by the widget?
                    # No, let the user override if needed, but the widget will update it.
                    pass
                
            widget.grid(row=i//4*2+1, column=i%4, sticky='ew', padx=5, pady=2)
            setattr(self, attr_name, widget)
            self.budget_widgets[attr_name] = widget

        def _on_budget_site_selected(e):
            # [FIX] 현장명을 가장 먼저 캡처 - focus_set() 호출 전에 해야 값이 유지됨
            selected_site = self.cb_budget_site.get().strip()
            # 폼 로드 (현장명 먼저 캡처했으므로 안전)
            self._load_budget_to_form(selected_site, silent=True)
            # 포커스 이동은 폼 로드 완료 후에 (after로 지연)
            self.root.after(50, self.ent_budget_revenue.focus_set)
        self.cb_budget_site.bind('<<ComboboxSelected>>', _on_budget_site_selected)

        # 2. Detailed Labor Cost Widget
        labor_detail_frame = ttk.LabelFrame(form_scrollable, text="인건비 상세 (Labor Cost Detail)", padding=10)
        labor_detail_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        def on_labor_change(total):
            self.ent_budget_labor.delete(0, tk.END)
            self.ent_budget_labor.insert(0, f"{total:,.0f}")
            self._update_budget_kpis()
            
        self.labor_detail_widget = LaborCostDetailWidget(labor_detail_frame, on_change_callback=on_labor_change)
        self.labor_detail_widget.pack(fill='x', expand=True)

        # 3. Detailed Material Cost Widget
        material_detail_frame = ttk.LabelFrame(form_scrollable, text="재료비 상세 (Material Cost Detail)", padding=10)
        material_detail_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        def on_material_change(total):
            self.ent_budget_material.delete(0, tk.END)
            self.ent_budget_material.insert(0, f"{total:,.0f}")
            self._update_budget_kpis()
            
        self.material_detail_widget = MaterialCostDetailWidget(material_detail_frame, on_change_callback=on_material_change)
        self.material_detail_widget.pack(fill='x', expand=True)

        # 4. Detailed Expense & Profit Widget
        expense_detail_frame = ttk.LabelFrame(form_scrollable, text="경비 및 이익 상세 (Expense & Profit Detail)", padding=10)
        expense_detail_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        def on_expense_change(exp_total, outsource_total, op_profit):
            self.ent_budget_expense.delete(0, tk.END)
            self.ent_budget_expense.insert(0, f"{exp_total:,.0f}")
            self.ent_budget_outsource.delete(0, tk.END)
            self.ent_budget_outsource.insert(0, f"{outsource_total:,.0f}")
            # [STABILITY] Update profit field if it exists
            if hasattr(self, 'ent_budget_profit'):
                # Calculate margin
                rev = get_rev()
                margin = (op_profit / rev * 100) if rev > 0 else 0.0
                self.ent_budget_profit.delete(0, tk.END)
                self.ent_budget_profit.insert(0, f"{op_profit:,.0f} ({margin:.1f}%)")
            self._update_budget_kpis()

        def get_lab():
            try: return float(self.ent_budget_labor.get().replace(',', '') or 0)
            except: return 0.0
        def get_mat():
            try: return float(self.ent_budget_material.get().replace(',', '') or 0)
            except: return 0.0
        def get_rev():
            try: 
                rev = float(self.ent_budget_revenue.get().replace(',', '') or 0)
                if rev == 0 and hasattr(self, 'ent_budget_unit_price'):
                    rev = float(self.ent_budget_unit_price.get().replace(',', '') or 0)
                return rev
            except: return 0.0
            
        self.expense_detail_widget = ExpenseProfitDetailWidget(
            expense_detail_frame, 
            on_change_callback=on_expense_change,
            get_labor_total_func=get_lab,
            get_material_total_func=get_mat,
            get_revenue_func=get_rev
        )
        self.expense_detail_widget.pack(fill='x', expand=True)
        
        # [NEW] 도급액 실시간 입력에 따른 이익금액 재계산 바인딩
        if hasattr(self, 'ent_budget_revenue'):
            self.ent_budget_revenue.bind('<KeyRelease>', lambda e: self.expense_detail_widget.calculate_all())

        # [MOVED to top_btn_frame]

        # ===============================================================
        # --- 하단: 현장별 일일사용량 조회 프레임 ---
        # ===============================================================
        bottom_container = ttk.Frame(main_paned)
        main_paned.add(bottom_container, weight=3)

        # 필터 영역
        bottom_filter = ttk.Frame(bottom_container)
        bottom_filter.pack(fill='x', pady=(5, 2))

        ttk.Label(bottom_filter, text="현장별 실적 조회", font=('Malgun Gothic', 10, 'bold')).pack(side='left', padx=(0, 15))

        ttk.Label(bottom_filter, text="현장:").pack(side='left', padx=5)
        self.cb_budget_view_site = ttk.Combobox(bottom_filter, width=20, state='readonly')
        site_values = ['전체'] + sorted([str(s).strip() for s in (self.sites or []) if str(s).strip()])
        self.cb_budget_view_site['values'] = site_values
        self.cb_budget_view_site.pack(side='left', padx=5)
        self.cb_budget_view_site.set('전체')
        self.cb_budget_view_site.bind('<<ComboboxSelected>>', lambda e: self.update_budget_site_view())

        ttk.Label(bottom_filter, text="시작일:").pack(side='left', padx=(10, 2))
        self.budget_view_start = DateEntry(bottom_filter, width=12, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        self.budget_view_start.pack(side='left', padx=2)
        today = datetime.datetime.now()
        self.budget_view_start.set_date(datetime.date(2024, 1, 1))

        ttk.Label(bottom_filter, text="종료일:").pack(side='left', padx=(5, 2))
        self.budget_view_end = DateEntry(bottom_filter, width=12, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        self.budget_view_end.pack(side='left', padx=2)
        self.budget_view_end.set_date(today)

        ttk.Button(bottom_filter, text="조회", command=self.update_budget_site_view).pack(side='left', padx=8)
        ttk.Button(bottom_filter, text="예산 수정/불러오기",
                   command=lambda: self._load_budget_to_form(self.cb_budget_view_site.get())).pack(side='left', padx=4)
        ttk.Button(bottom_filter, text="컬럼 추가", command=self.add_budget_view_custom_column).pack(side='left', padx=4)
        ttk.Button(bottom_filter, text="컬럼 관리", command=self.show_budget_view_column_dialog).pack(side='left', padx=4)
        ttk.Button(bottom_filter, text="엑셀 내보내기", command=self.export_budget_sales_status).pack(side='right', padx=10)

        # KPI 요약 행 제거 (예산입력 5번 영업이익률로 대체)
        # Treeview (일일사용량 코럼)
        tree_outer = ttk.Frame(bottom_container)
        tree_outer.pack(fill='both', expand=True)

        # 전체 컬럼: 기본 표시 컬럼 + 사용자가 "추가" 할 수 있는 숨김 기본 컬럼 + 사용자정의 컬럼
        self.budget_view_builtin_cols = (
            'Date', 'Site', '장비명', '검사방법', '검사량', '단가', '검사단가', '출장비', '일식',
            'OT합계', '침투제', '세척제', '현상제', '자재단가', '합계', '비고',
            '작업자', '품목명', '입력시간', '차량번호', '주행거리', '차량점검', '차량비고',
            'MaterialID'
        )
        self.budget_view_builtin_width_map = {
            'Date': 90, 'Site': 120, '장비명': 100, '검사방법': 80, '검사량': 70, '단가': 80,
            '검사단가': 90, '출장비': 80, '일식': 70, 'OT합계': 80, '침투제': 70, '세척제': 70, '현상제': 70, '자재단가': 90,
            '합계': 100, '비고': 150, '작업자': 170, '품목명': 180, '입력시간': 140,
            '차량번호': 110, '주행거리': 100, '차량점검': 120, '차량비고': 140, 'MaterialID': 90
        }
        self.budget_view_builtin_head_map = {
            'Date': '날짜', 'Site': '현장', '장비명': '장비', '검사방법': '검사방법', '검사량': '수량', '단가': '단가',
            '검사단가': '매출금액', '출장비': '출장비', '일식': '일식', 'OT합계': 'OT합계', '침투제': '침투제', '세척제': '세척제', '현상제': '현상제',
            '자재단가': '자재단가', '합계': '합계', '비고': '비고', '작업자': '작업자', '품목명': '품목명',
            '입력시간': '입력시간', '차량번호': '차량번호', '주행거리': '주행거리', '차량점검': '차량점검',
            '차량비고': '차량비고', 'MaterialID': '자재ID'
        }
        self.budget_view_heading_aliases = getattr(self, 'budget_view_heading_aliases', {})
        self.budget_view_custom_columns = getattr(self, 'budget_view_custom_columns', [])
        self.budget_view_default_cols = (
            'Date', 'Site', '장비명', '검사방법', '검사량', '단가', '검사단가', '출장비', '일식',
            'OT합계', '침투제', '세척제', '현상제', '자재단가', '합계', '비고'
        )

        vsb = ttk.Scrollbar(tree_outer, orient='vertical')
        hsb = ttk.Scrollbar(tree_outer, orient='horizontal')
        self.budget_view_tree = ttk.Treeview(tree_outer, columns=self.budget_view_builtin_cols, show='headings',
                                              yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._refresh_budget_view_tree_columns(reload_data=False)

        vsb.config(command=self.budget_view_tree.yview)
        hsb.config(command=self.budget_view_tree.xview)
        self.budget_view_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        tree_outer.grid_rowconfigure(0, weight=1)
        tree_outer.grid_columnconfigure(0, weight=1)
        self.budget_view_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())
        self.enable_tree_column_drag(self.budget_view_tree, context_menu_handler=self._show_budget_view_heading_context_menu)

        # ── ESC: 현장별 탭 내 모든 입력 위젯에서 포커스 해제 ──────────────
        def _esc_to_root(e=None):
            self.root.focus_set()
            return "break"

        def _bind_esc_recursive(widget):
            try:
                cls = widget.winfo_class()
                # Toplevel(달력 팝업 등) 내부는 절대 순회하지 않음
                # → 순회만으로도 winfo_children() 가 Toplevel을 화면에 표시시킬 수 있음
                if cls == 'Toplevel':
                    return
                if cls in ('TEntry', 'Entry', 'TCombobox'):
                    widget.bind('<Escape>', _esc_to_root, add='+')
            except Exception:
                pass
            for child in widget.winfo_children():
                _bind_esc_recursive(child)

        # after_idle: 모든 위젯 생성 완료 후 바인딩
        self.root.after_idle(lambda: _bind_esc_recursive(self.tab_daily_usage))
        
        # 초기 데이터 로드 (첫 현장 자동 선택 후 표시)
        self.root.after(100, lambda: self.update_budget_site_view())

    def update_budget_view(self):
        """Refresh the budget inquiry view to reflect updated budget data."""
        if hasattr(self, 'cb_budget_view_site'):
            # Only refresh if a site is actually being viewed
            if self.cb_budget_view_site.get():
                self.update_budget_site_view()

    def _refresh_budget_view_tree_columns(self, reload_data=False):
        """Rebuild budget site performance tree column structure."""
        if not hasattr(self, 'budget_view_tree'):
            return

        custom_cols = getattr(self, 'budget_view_custom_columns', []) or []
        # [FIX] built-in 컬럼과 이름/키가 중복되는 사용자 컬럼 정리
        builtin_keys = set(getattr(self, 'budget_view_builtin_cols', []) or [])
        builtin_headings = set((getattr(self, 'budget_view_builtin_head_map', {}) or {}).values())
        sanitized_custom_cols = []
        seen_custom_keys = set()
        seen_custom_names = set()
        for c in custom_cols:
            key = str(c.get('key', '')).strip()
            name = str(c.get('name', key)).strip()
            if not key:
                continue
            # built-in 과 키 충돌/이름 충돌이면 제거
            if key in builtin_keys or name in builtin_headings:
                continue
            # custom 내부 중복도 제거
            if key in seen_custom_keys or name in seen_custom_names:
                continue
            seen_custom_keys.add(key)
            seen_custom_names.add(name)
            sanitized_custom_cols.append(c)
        if len(sanitized_custom_cols) != len(custom_cols):
            self.budget_view_custom_columns = sanitized_custom_cols
            custom_cols = sanitized_custom_cols

        custom_keys = [c.get('key') for c in custom_cols if c.get('key')]
        self.budget_view_cols = tuple(list(self.budget_view_builtin_cols) + custom_keys)

        head_map = dict(getattr(self, 'budget_view_builtin_head_map', {}))
        width_map = dict(getattr(self, 'budget_view_builtin_width_map', {}))
        for col in custom_cols:
            key = col.get('key')
            if not key:
                continue
            head_map[key] = col.get('name', key)
            width_map[key] = int(col.get('width', 120) or 120)

        alias_map = getattr(self, 'budget_view_heading_aliases', {}) or {}
        self.budget_view_tree.configure(columns=self.budget_view_cols)
        for col in self.budget_view_cols:
            heading_text = alias_map.get(col, head_map.get(col, col))
            self.budget_view_tree.heading(col, text=heading_text,
                                          command=lambda c=col: self.treeview_sort_column(self.budget_view_tree, c, False))
            self.budget_view_tree.column(col, width=width_map.get(col, 100), minwidth=40, anchor='center')

        visible_source = getattr(self, 'budget_view_visible_cols', []) or list(self.budget_view_default_cols)
        visible = [c for c in visible_source if c in self.budget_view_cols]
        self.budget_view_tree['displaycolumns'] = visible if visible else self.budget_view_default_cols
        self.budget_view_visible_cols = list(self.budget_view_tree['displaycolumns'])

        if reload_data:
            self.update_budget_site_view()

    def add_budget_view_custom_column(self):
        """Add a custom column to budget site performance view."""
        name = simpledialog.askstring("컬럼 추가", "새 컬럼 이름을 입력하세요:", parent=self.root)
        if not name or not str(name).strip():
            return
        name = str(name).strip()

        # [FIX] built-in/기존 사용자 컬럼과 이름 중복 방지
        builtin_names = set((getattr(self, 'budget_view_builtin_head_map', {}) or {}).values())
        existing_custom_names = {str(c.get('name', '')).strip() for c in (getattr(self, 'budget_view_custom_columns', []) or [])}
        if name in builtin_names or name in existing_custom_names:
            messagebox.showwarning("중복 컬럼", f"'{name}' 컬럼은 이미 존재합니다.")
            return

        default_value = simpledialog.askstring("컬럼 내용", f"'{name}' 컬럼의 기본 내용을 입력하세요(선택):", parent=self.root)

        existing = set(getattr(self, 'budget_view_cols', []))
        key_base = re.sub(r'[^0-9A-Za-z가-힣_]+', '_', name).strip('_') or '사용자컬럼'
        key = key_base
        idx = 1
        while key in existing:
            idx += 1
            key = f"{key_base}_{idx}"

        custom_cols = getattr(self, 'budget_view_custom_columns', []) or []
        custom_cols.append({'key': key, 'name': name, 'default': default_value or '', 'width': 120})
        self.budget_view_custom_columns = custom_cols

        visible = list(getattr(self, 'budget_view_visible_cols', []) or list(self.budget_view_default_cols))
        if key not in visible:
            visible.append(key)
        self.budget_view_visible_cols = visible

        self._refresh_budget_view_tree_columns(reload_data=True)
        self.save_tab_config()

    def rename_budget_view_column(self, col_key):
        """Rename a budget site performance column heading."""
        if not col_key:
            return
        current_name = self.budget_view_tree.heading(col_key)['text'] if hasattr(self, 'budget_view_tree') else col_key
        new_name = simpledialog.askstring("컬럼 이름 변경", "새 컬럼 이름을 입력하세요:", initialvalue=current_name, parent=self.root)
        if not new_name or not str(new_name).strip():
            return
        new_name = str(new_name).strip()

        # Custom column: rename the stored name itself
        updated = False
        for col in getattr(self, 'budget_view_custom_columns', []) or []:
            if col.get('key') == col_key:
                col['name'] = new_name
                updated = True
                break
        if not updated:
            aliases = getattr(self, 'budget_view_heading_aliases', {}) or {}
            default_name = getattr(self, 'budget_view_builtin_head_map', {}).get(col_key, col_key)
            if new_name == default_name:
                aliases.pop(col_key, None)
            else:
                aliases[col_key] = new_name
            self.budget_view_heading_aliases = aliases

        self._refresh_budget_view_tree_columns(reload_data=False)
        self.save_tab_config()

    def set_budget_view_custom_column_content(self, col_key):
        """Change default content for a custom column."""
        custom_cols = getattr(self, 'budget_view_custom_columns', []) or []
        target = next((c for c in custom_cols if c.get('key') == col_key), None)
        if not target:
            return
        new_value = simpledialog.askstring("컬럼 내용 설정", f"'{target.get('name', col_key)}' 컬럼의 기본 내용을 입력하세요:",
                                           initialvalue=target.get('default', ''), parent=self.root)
        if new_value is None:
            return
        target['default'] = new_value
        self._refresh_budget_view_tree_columns(reload_data=True)
        self.save_tab_config()

    def delete_budget_view_custom_column(self, col_key):
        """Delete a custom budget column."""
        custom_cols = getattr(self, 'budget_view_custom_columns', []) or []
        target = next((c for c in custom_cols if c.get('key') == col_key), None)
        if not target:
            return
        if not messagebox.askyesno("컬럼 삭제", f"'{target.get('name', col_key)}' 컬럼을 삭제하시겠습니까?", parent=self.root):
            return

        self.budget_view_custom_columns = [c for c in custom_cols if c.get('key') != col_key]
        self.budget_view_visible_cols = [c for c in getattr(self, 'budget_view_visible_cols', []) if c != col_key]
        aliases = getattr(self, 'budget_view_heading_aliases', {}) or {}
        aliases.pop(col_key, None)
        self.budget_view_heading_aliases = aliases
        self._refresh_budget_view_tree_columns(reload_data=True)
        self.save_tab_config()

    def show_budget_view_column_dialog(self):
        """Open dialog to show/hide columns in budget site performance view"""
        if not hasattr(self, 'budget_view_tree'):
            return

        all_cols = list(self.budget_view_tree['columns'])
        active_cols = self.budget_view_tree['displaycolumns']
        if not active_cols or active_cols == ('#all'):
            active_cols = all_cols

        dialog_cols = [(col, self.budget_view_tree.heading(col)['text']) for col in all_cols]
        dialog = ColumnSelectionDialog(self.root, dialog_cols, title="현장별 실적 표시 컬럼 관리")
        for col, var in dialog.vars.items():
            var.set(col in active_cols)

        dialog.wait_window()

        if dialog.result is not None:
            # 필수 컬럼 보정: 날짜/현장은 항상 유지
            final_selection = list(dialog.result)
            for mandatory in ('Date', 'Site'):
                if mandatory in all_cols and mandatory not in final_selection:
                    final_selection.append(mandatory)

            sorted_selection = [c for c in all_cols if c in final_selection]
            self.budget_view_tree['displaycolumns'] = sorted_selection if sorted_selection else all_cols
            self.budget_view_visible_cols = list(self.budget_view_tree['displaycolumns'])
            self.save_tab_config()

    def _show_budget_view_heading_context_menu(self, event):
        """공사탭 현장별 실적 Treeview 헤더 우클릭 메뉴"""
        if not hasattr(self, 'budget_view_tree'):
            return
        tree = self.budget_view_tree
        try:
            if tree.identify_region(event.x, event.y) != 'heading':
                return
        except Exception:
            return

        column_id = tree.identify_column(event.x)
        col_key = self._get_column_name_from_id(tree, column_id)
        if not col_key:
            return
        col_text = tree.heading(col_key)['text']
        custom_keys = {c.get('key') for c in (getattr(self, 'budget_view_custom_columns', []) or [])}

        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label=f"⬅️ '{col_text}' 왼쪽으로 이동", command=lambda: self._move_column_visual(tree, column_id, -1))
        menu.add_command(label=f"➡️ '{col_text}' 오른쪽으로 이동", command=lambda: self._move_column_visual(tree, column_id, 1))
        menu.add_separator()
        menu.add_command(label=f"✏️ '{col_text}' 이름 변경...", command=lambda: self.rename_budget_view_column(col_key))
        menu.add_command(label="➕ 사용자 컬럼 추가...", command=self.add_budget_view_custom_column)
        if col_key in custom_keys:
            menu.add_command(label=f"📝 '{col_text}' 내용 설정...", command=lambda: self.set_budget_view_custom_column_content(col_key))
            menu.add_command(label=f"🗑️ '{col_text}' 사용자 컬럼 삭제", command=lambda: self.delete_budget_view_custom_column(col_key))
        menu.add_separator()
        menu.add_command(label="⚙️ 컬럼 관리(추가/삭제)...", command=self.show_budget_view_column_dialog)
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def update_budget_site_view(self):
        """현장별 일일사용량 데이터를 하단 Treeview에 표시하고 KPI를 갱신한다."""
        try:
            if not hasattr(self, 'budget_view_tree'):
                return

            for item in self.budget_view_tree.get_children():
                self.budget_view_tree.delete(item)

            if not hasattr(self, 'cb_budget_view_site'):
                return
            site = self.cb_budget_view_site.get().strip()
            
            # [NEW] Track logically unique entries to avoid double-counting materials from split-row records
            processed_entry_ids = set()

            # [FIX] Do NOT return early if df is empty, we still need to run column hiding logic
            if self.daily_usage_df.empty:
                df_empty = True
            else:
                df_empty = False

            # --- 날짜 필터 ---
            start_date = self.budget_view_start.get_date()
            end_date   = self.budget_view_end.get_date()
            df = self.daily_usage_df.copy()
            # 구버전 컬럼명 호환
            if 'Site' not in df.columns and '현장' in df.columns:
                df['Site'] = df['현장']
            if 'Date' not in df.columns and '날짜' in df.columns:
                df['Date'] = df['날짜']
            if 'Site' not in df.columns:
                df_empty = True
            
            if not df_empty:
                df['_site_norm'] = df['Site'].astype(str).str.strip()

            # 날짜 파싱(실패 데이터는 제외 대신 전체 누락 방지를 위해 후단에서 완화)
            if 'Date' in df.columns:
                # Use normalized Timestamps for robust comparison with tk date objects
                df['_date'] = pd.to_datetime(df['Date'], errors='coerce').dt.normalize()
                start_ts = pd.to_datetime(start_date).normalize()
                end_ts = pd.to_datetime(end_date).normalize()
                date_mask = (df['_date'] >= start_ts) & (df['_date'] <= end_ts)
                # [STRICT] Exclude any rows with invalid/missing dates (NaT)
                date_mask = date_mask & df['_date'].notna()
            else:
                date_mask = pd.Series([True] * len(df), index=df.index)

            if site and site != '전체':
                site_mask = (df['_site_norm'] == site)
            else:
                site_mask = pd.Series([True] * len(df), index=df.index)

            mask = date_mask & site_mask
            df = df[mask].copy()

            # --- 자재 원가 맵 ---
            mat_id_cost_map = {}
            if 'MaterialID' in self.materials_df.columns and '원가' in self.materials_df.columns:
                mat_id_cost_map = self.materials_df.set_index('MaterialID')['원가'].fillna(0).to_dict()

            def _f(val):
                if pd.isna(val) or str(val).lower() == 'nan': return 0.0
                try: 
                    s = str(val).replace(',', '').strip()
                    if not s: return 0.0
                    num = float(s)
                    return 0.0 if np.isnan(num) else num
                except: return 0.0

            total_income = 0.0   # 검사비 + OT 합계
            total_expense = 0.0  # 출장비 + 일식
            total_net_revenue = 0.0 # 검사비 합계
            total_sum     = 0.0
            total_mat_cost = 0.0 # 자재 실적 사용료 합계
            total_ndt_penetrant = 0.0 # NDT 침투제 합계
            total_ndt_cleaner = 0.0 # NDT 세척제 합계
            total_ndt_developer = 0.0 # NDT 현상제 합계

            import re as _re
            
            # [NEW] Robust is_active helper for smart column hiding
            def is_active(val):
                if val is None: return False
                s = str(val).strip().lower()
                if s in ('', '0', '0.0', '0.00', 'nan', 'none', '-', '0원', '0.0원', '0시간'):
                    return False
                try:
                    # Remove common units for numeric check
                    v_raw = s.replace(',', '').replace('원', '').replace('시간', '').strip()
                    if not v_raw: return False
                    v = float(v_raw)
                    return abs(v) > 0.001
                except:
                    return bool(s)

            # [NEW] Smart Column Hiding tracking
            mandatory_cols = {'Date', 'Site', '합계', '비고'}
            has_data_map = {col: False for col in self.budget_view_cols if col not in mandatory_cols}

            def _clean_str(val):
                if pd.isna(val):
                    return ''
                s = str(val).strip()
                return '' if s.lower() in ('nan', 'none') else s

            def _first_text(_row, keys, default=''):
                """여러 후보 컬럼 중 첫 유효 문자열 반환 (구버전 컬럼명 호환)."""
                for k in keys:
                    if k in _row.index:
                        v = _clean_str(_row.get(k, ''))
                        if v:
                            return v
                return default

            def _ndt_val(_row, *keys):
                for k in keys:
                    if k in _row.index:
                        return _f(_row.get(k, 0))
                return 0.0

            for _, row in df.iterrows():
                # [NEW] Logical entry deduplication for the display list (Treeview)
                e_date = str(row.get('Date', '')).split(' ')[0]
                e_site = str(row.get('Site', '')).strip()
                raw_time = str(row.get('EntryTime', '')).strip()
                e_time = raw_time[:16] if len(raw_time) > 16 else raw_time
                e_uid = f"{e_date}|{e_site}|{e_time}"
                
                if e_uid in processed_entry_ids:
                    continue # Skip redundant rows from being displayed in the list
                processed_entry_ids.add(e_uid)

                # [FIX] OT 컬럼에서 금액 추출 로직 강화 (NaN 방지)
                parsed_ots = []
                for i in range(1, 11):
                    raw_ot = row.get('OT' if i == 1 else f'OT{i}', 0)
                    if pd.isna(raw_ot) or not str(raw_ot).strip() or str(raw_ot).lower() == 'nan':
                        parsed_ots.append(0.0)
                        continue
                    
                    v_str = str(raw_ot).strip()
                    if '(' in v_str and '원)' in v_str:
                        try:
                            amt_str = v_str.split('(')[1].split('원')[0].replace(',', '').strip()
                            parsed_ots.append(_f(amt_str))
                        except: parsed_ots.append(0.0)
                    else:
                        parsed_ots.append(_f(v_str))
                
                ot_sum = sum(parsed_ots)

                net    = _f(row.get('검사비', 0))
                travel = _f(row.get('출장비', 0))
                meal   = _f(row.get('일식', 0))
                qty    = _f(row.get('검사량', 0))
                price  = _f(row.get('단가', 0))

                # 자재원가
                test_method = str(row.get('검사방법', '')).upper()
                if 'RT' in test_method:
                    usage = _f(row.get('Usage', 0))
                else:
                    usage = _f(row.get('Usage', 0))

                # [NEW] Track logically unique entries to avoid double-counting materials from split-row records
                # Based on Date + Site + EntryTime (best UID) or Date + Site + Note (as fallback)
                e_time = str(row.get('EntryTime', '')).strip()
                e_site = str(row.get('Site', row.get('현장', ''))).strip()
                e_date = str(row.get('Date', row.get('날짜', ''))).strip()
                e_note = str(row.get('Note', row.get('비고', ''))).strip()
                e_uid = f"{e_date}|{e_site}|{e_time}|{e_note}"
                
                is_new_logical_entry = (e_uid not in processed_entry_ids)
                processed_entry_ids.add(e_uid)

                # 일일 탭 NDT 입력값(침투제/세척제/현상제)
                # [FIX] ONLY sum consumables for the first row of each logical entry to prevent inflation
                ndt_penetrant = 0.0
                ndt_cleaner = 0.0
                ndt_developer = 0.0
                
                if is_new_logical_entry:
                    ndt_penetrant = _ndt_val(row, 'NDT_침투제', '침투제', 'NDT_침투액', '침투액')
                    ndt_cleaner = _ndt_val(row, 'NDT_세척제', '세척제', 'NDT_세척액', '세척액')
                    ndt_developer = _ndt_val(row, 'NDT_현상제', '현상제', 'NDT_현상액', '현상액')

                    total_ndt_penetrant += ndt_penetrant
                    total_ndt_cleaner += ndt_cleaner
                    total_ndt_developer += ndt_developer

                unit_cost = mat_id_cost_map.get(row.get('MaterialID'), 0)
                mat_cost = usage * float(unit_cost)
                # Material cost should also be counted carefully (main items are row-wise, but consumables are entry-wise)
                # However, for consistency with the material cost section, we track total_mat_cost here.
                total_mat_cost += mat_cost

                # --- Financial Totals (Sum for ALL rows) ---
                total_net_revenue += net
                total_expense += (travel + meal)
                total_income += (net + ot_sum)
                total_sum += (net + travel + meal)

                # 작업자 통합
                worker_cols = ['User'] + [f'User{i}' for i in range(2, 11)]
                raw_workers = []
                for wcol in worker_cols:
                    wv = _clean_str(row.get(wcol, ''))
                    if wv and wv not in raw_workers:
                        raw_workers.append(wv)
                consolidated_workers = ', '.join(raw_workers)

                # 품목명 표시명
                try:
                    mat_name = self.get_material_display_name(row.get('MaterialID')) if pd.notna(row.get('MaterialID')) else ''
                except Exception:
                    mat_name = ''

                row_map = {
                    'Date': _clean_str(row.get('Date', '')),
                    'Site': _clean_str(row.get('Site', '')),
                    '장비명': _clean_str(row.get('장비명', '')),
                    '검사방법': _clean_str(row.get('검사방법', '')),
                    '검사량': f"{qty:g}",
                    '단가': f"{price:,.0f}",
                    '검사단가': f"{net:,.0f}",
                    '출장비': f"{travel:,.0f}",
                    '일식': f"{meal:,.0f}",
                    'OT합계': f"{ot_sum:,.0f}",
                    '침투제': f"{_f(row.get('NDT_침투제', row.get('침투제', 0))):g}",
                    '세척제': f"{_f(row.get('NDT_세척제', row.get('세척제', 0))):g}",
                    '현상제': f"{_f(row.get('NDT_현상제', row.get('현상제', 0))):g}",
                    '자재단가': f"{unit_cost:,.0f}",
                    '합계': f"{(net + travel + meal):,.0f}",
                    '비고': _clean_str(row.get('비고', row.get('Note', ''))),
                    '작업자': consolidated_workers,
                    '품목명': mat_name,
                    '입력시간': _clean_str(row.get('EntryTime', '')),
                    '차량번호': _first_text(row, ['차량번호', 'vehicle_info', 'VehicleInfo']),
                    '주행거리': _first_text(row, ['주행거리', 'mileage', 'Mileage']),
                    '차량점검': _first_text(row, ['차량점검', '차량 점검', 'vehicle_checks', 'inspection', 'Inspection']),
                    '차량비고': _first_text(row, ['차량비고', '차량 비고', 'remarks', 'vehicle_remarks', 'Remark']),
                    'MaterialID': _clean_str(row.get('MaterialID', '')),
                    'Usage': _clean_str(row.get('Usage', '')),
                }
                for custom_col in getattr(self, 'budget_view_custom_columns', []) or []:
                    key = custom_col.get('key')
                    if key:
                        row_map[key] = custom_col.get('default', '')
                
                # [NEW] Mark columns that have non-zero/non-empty data using robust is_active
                for col, val in row_map.items():
                    if col in has_data_map and not has_data_map[col]:
                        if is_active(val):
                            has_data_map[col] = True

                self.budget_view_tree.insert('', tk.END, values=tuple(row_map.get(col, '') for col in self.budget_view_cols))

                try:
                    equip_name = str(row.get('장비명', ''))
                    m_name_local = str(row.get('검사방법', ''))
                    date_key = str(row.get('Date', ''))
                    if 'RT' in m_name_local or 'RT' in equip_name or 'RT' in mat_name:
                        rt_dates.add(date_key)
                    
                    # [NEW] Generic equipment tracking by name
                    if equip_name and equip_name.lower() != 'nan':
                        if equip_name not in equip_dates_map:
                            equip_dates_map[equip_name] = set()
                        equip_dates_map[equip_name].add(date_key)
                except Exception:
                    pass

            # --- 합계 행 추가 ---
            if df.shape[0] > 0:
                total_row_map = {
                    'Date': f"[합계]",
                    'Site': f"{site}",
                    '장비명': '',
                    '검사방법': '',
                    '검사량': '',
                    '단가': '',
                    '검사단가': f"{total_net_revenue:,.0f}",
                    '출장비': f"{total_expense:,.0f}",
                    '일식': '',
                    'OT합계': '',
                    '자재사용량': '',
                    '침투제': f"{total_ndt_penetrant:g}",
                    '세척제': f"{total_ndt_cleaner:g}",
                    '현상제': f"{total_ndt_developer:g}",
                    '자재단가': '',
                    '합계': f"{total_sum:,.0f}",
                    '비고': f"소계: 수입 {total_income:,.0f} 경비 {total_expense:,.0f} 자재원가 {total_mat_cost:,.0f}",
                    '작업자': '',
                    '품목명': '',
                    '입력시간': '',
                    '차량번호': '',
                    '주행거리': '',
                    '차량점검': '',
                    '차량비고': '',
                    'MaterialID': '',
                    'Usage': '',
                }
                for custom_col in getattr(self, 'budget_view_custom_columns', []) or []:
                    key = custom_col.get('key')
                    if key:
                        total_row_map[key] = ''
                self.budget_view_tree.insert('', tk.END, values=tuple(total_row_map.get(col, '') for col in self.budget_view_cols), tags=('total',))

            # --- [NEW] 예산 대비 실적 요약 행 추가 ---
            if df.shape[0] > 0 and site and site != '전체' and hasattr(self, 'budget_df') and not self.budget_df.empty:
                budget_match = self.budget_df[self.budget_df['Site'] == site]
                if not budget_match.empty:
                    b_row = budget_match.iloc[0]
                    
                    # Fetch Budget Values
                    def to_f_b(key):
                        try: 
                            v = str(b_row.get(key, 0)).replace(',', '').strip()
                            return float(v) if v else 0.0
                        except: return 0.0
                    
                    b_rev = to_f_b('Revenue')
                    b_labor = to_f_b('LaborCost')
                    b_mat = to_f_b('MaterialCost')
                    b_exp = to_f_b('Expense')
                    b_profit = to_f_b('Profit')
                    
                    # Progress / Variance Calculations
                    rev_progress = (total_income / b_rev * 100) if b_rev > 0 else 0
                    mat_progress = (total_mat_cost / b_mat * 100) if b_mat > 0 else 0
                    exp_progress = (total_expense / b_exp * 100) if b_exp > 0 else 0
                    
                    # Final Net Profit Achievement
                    actual_profit = total_income - total_expense - total_mat_cost
                    
                    comparison_row_map = {
                        'Date': f"[예산대비]",
                        'Site': f"달성률: {rev_progress:.1f}%",
                        '장비명': '',
                        '검사방법': '',
                        '검사량': '',
                        '단가': '',
                        '검사단가': f"예산 {b_rev:,.0f}",
                        '출장비': f"경비 {exp_progress:.1f}%",
                        '일식': '',
                        'OT합계': '',
                        '자재사용량': f"자재 {mat_progress:.1f}%",
                        '침투제': '',
                        '세척제': '',
                        '현상제': '',
                        '자재단가': '',
                        '합계': f"{actual_profit:,.0f}",
                        '비고': f"목표이익: {b_profit:,.0f} | 현재이익: {actual_profit:,.0f} ({actual_profit - b_profit:+,.0f})",
                        '작업자': '',
                        '품목명': '',
                        '입력시간': '',
                        '차량번호': '',
                        '주행거리': '',
                        '차량점검': '',
                        '차량비고': '',
                        'MaterialID': '',
                        'Usage': '',
                    }
                    for custom_col in getattr(self, 'budget_view_custom_columns', []) or []:
                        key = custom_col.get('key')
                        if key:
                            comparison_row_map[key] = ''
                    
                    # Insert with a distinct tag for styling
                    self.budget_view_tree.tag_configure('comparison', background='#E3F2FD', font=('Malgun Gothic', 10, 'bold'))
                    self.budget_view_tree.insert('', tk.END, values=tuple(comparison_row_map.get(col, '') for col in self.budget_view_cols), tags=('comparison',))

            # --- [NEW] Smart Column Hiding Application ---
            active_cols = [col for col in self.budget_view_cols 
                           if col in mandatory_cols or has_data_map.get(col, False)]
            self.budget_view_tree['displaycolumns'] = active_cols
            
            # --- Treeview 태그 스타일 (합계 행 강조) ---

            if not self.budget_view_tree.tag_configure('total'):
                self.budget_view_tree.tag_configure('total', background='#ffff00', foreground='#000000')
        except Exception as e:
            print(f"ERROR in update_budget_site_view: {e}")


    def _update_budget_kpis(self):
        """Update the top KPI summary labels based on current budget form values"""
        try:
            def _get_val(widget):
                try: return float(str(widget.get()).replace(',', '').split(' ')[0] or 0)
                except: return 0.0
                
            rev = _get_val(self.ent_budget_revenue)
            if rev == 0 and hasattr(self, 'ent_budget_unit_price'):
                rev = _get_val(self.ent_budget_unit_price) # Fallback to unit price if revenue empty
                
            lab = _get_val(self.ent_budget_labor)
            mat = _get_val(self.ent_budget_material)
            exp = _get_val(self.ent_budget_expense)
            out = _get_val(self.ent_budget_outsource)
            
            # 4. Total Cost (Add Indirect Costs if needed, but here we sync from widget)
            # [NEW] As per user request, KPI "실행원가" now shows Grand Total Cost (including Indirect Costs)
            if hasattr(self, 'expense_detail_widget') and hasattr(self.expense_detail_widget, 'get_total_cost'):
                # The widget's total already includes insurance/depreciation/indirect costs in sections 1~7
                # but if we want exactly the same logic as the widget:
                try: 
                    # Get final total from the widget's label to match "4. 총원가" perfectly
                    raw_total_text = self.expense_detail_widget.lbl_grand_total_cost.cget('text')
                    total_cost = float("".join(c for c in raw_total_text if c.isdigit() or c == '.') or 0)
                except:
                    total_cost = lab + mat + exp + out
            else:
                total_cost = lab + mat + exp + out
                
            profit = rev - total_cost
            
            # [FIX] Safety check for Margin to prevent NaN or Inf
            if rev > 0:
                margin = (profit / rev * 100)
            else:
                margin = 0.0
            
            if np.isnan(profit) or np.isinf(profit): profit = 0.0
            if np.isnan(margin) or np.isinf(margin): margin = 0.0

            
            if hasattr(self, 'lbl_kpi_rev'):
                self.lbl_kpi_rev.config(text=f"{rev:,.0f}원")
                self.lbl_kpi_cost.config(text=f"{total_cost:,.0f}원")
                self.lbl_kpi_profit.config(text=f"{profit:,.0f}원", foreground="#ef4444" if profit < 0 else "#10b981")
                self.lbl_kpi_margin.config(text=f"{margin:.1f}%", foreground="#ef4444" if margin < 0 else "#10b981")
                
            # [NEW] Sync calculated profit and margin to the top entry fields
            if hasattr(self, 'ent_budget_profit'):
                self.ent_budget_profit.delete(0, tk.END)
                self.ent_budget_profit.insert(0, f"{profit:,.0f}")
            if hasattr(self, 'ent_budget_margin'):
                self.ent_budget_margin.delete(0, tk.END)
                self.ent_budget_margin.insert(0, f"{margin:.1f}")
            
            # [STABILITY] Force UI refresh to make changes visible
            self.root.update_idletasks()
        except Exception as e:
            print(f"DEBUG: Budget KPI sync error: {e}")

    def _load_budget_to_form(self, site, silent=False):
        """선택한 현장의 예산서를 상단 입력 폼에 로드한다."""
        # [NEW] Clear current form before loading new site data to prevent residual data
        self.clear_budget_form()

        # [FIX] clear_budget_form()이 cb_budget_site를 지우므로 즉시 복원
        # budget_df에 해당 현장이 없더라도 콤보박스에 선택한 현장명은 유지해야 함
        if site and hasattr(self, 'cb_budget_site'):
            self.cb_budget_site.set(site)

        if not site:
            if not silent: messagebox.showwarning("현장 미선택", "현장을 먼저 선택해주세요.")
            return
        if not hasattr(self, 'budget_df') or self.budget_df.empty:
            if not silent: messagebox.showinfo("데이터 없음", f"'{site}' 현장의 저장된 예산서가 없습니다.")
            return
        match = self.budget_df[self.budget_df['Site'] == site]
        if match.empty:
            if not silent: messagebox.showinfo("데이터 없음", f"'{site}' 현장의 저장된 예산서가 없습니다.")
            return

        row = match.iloc[0]
        self.cb_budget_site.set(site)
        for attr, col in [('ent_budget_revenue',   'Revenue'),
                          ('ent_budget_unit_price','UnitPrice'),
                          ('ent_budget_labor',     'LaborCost'),
                          ('ent_budget_material',  'MaterialCost'),
                          ('ent_budget_expense',   'Expense'),
                          ('ent_budget_outsource', 'OutsourceCost'),
                          ('ent_budget_profit',    'Profit'),
                          ('ent_budget_note',      'Note')]:
            w = getattr(self, attr, None)
            if w:
                w.delete(0, tk.END)
                val = row.get(col, '')
                # [FIX] Sanitize NaN values to prevent "nan" string appearing in entry boxes
                if pd.isna(val) or str(val).lower() == 'nan':
                    display_val = "" if col == 'Note' else "0"
                else:
                    if col != 'Note' and col != 'Profit':
                        try:
                            # Re-sanitize and format to ensure float stability
                            f_val = float(str(val).replace(',', '').strip() or 0)
                            display_val = f"{f_val:,.0f}"
                        except:
                            display_val = str(val)
                    else:
                        display_val = str(val)
                w.insert(0, display_val)

        import json as _json
        for json_col, widget_attr in [('LaborDetail',   'labor_detail_widget'),
                                       ('MaterialDetail','material_detail_widget'),
                                       ('ExpenseDetail', 'expense_detail_widget')]:
            jdata = row.get(json_col, '')
            widget = getattr(self, widget_attr, None)
            if widget and jdata and str(jdata).strip():
                try:
                    widget.set_data(_json.loads(jdata))
                except:
                    widget.reset()
            elif widget:
                # [NEW] Clear data if no saved detail exists for this site
                widget.reset()

        # [FIX] set_data() → calculate_all() → on_change_callback() 순서로 금액 필드가
        # 상세 항목 합계로 덮어씌워지는 문제 방지.
        # 상세 위젯 로드 완료 후 저장된 원본 금액 값을 명시적으로 재설정한다.
        for attr, col in [('ent_budget_revenue',   'Revenue'),
                          ('ent_budget_unit_price','UnitPrice'),
                          ('ent_budget_labor',     'LaborCost'),
                          ('ent_budget_material',  'MaterialCost'),
                          ('ent_budget_expense',   'Expense'),
                          ('ent_budget_outsource', 'OutsourceCost'),
                          ('ent_budget_profit',    'Profit')]:
            w = getattr(self, attr, None)
            if w:
                val = row.get(col, '')
                try:
                    display = f"{float(val):,.0f}" if val else '0'
                    if attr == 'ent_budget_profit':
                        rev_val = row.get('Revenue', 0)
                        try: rev_float = float(rev_val)
                        except: rev_float = 0.0
                        prof_float = float(val) if val else 0.0
                        margin = (prof_float / rev_float * 100) if rev_float > 0 else 0.0
                        display = f"{prof_float:,.0f} ({margin:.1f}%)"
                except:
                    display = '0'
                w.delete(0, tk.END)
                w.insert(0, display)

        if not silent: messagebox.showinfo("로드 완료", f"'{site}' 현장의 예산서를 불러왔습니다.")
        # [STABILITY] Update KPI summary after load
        self._update_budget_kpis()

    def fill_budget_from_actuals(self):
        """하단 필터(현장, 기간) 기준으로 실제 실적 데이터를 집계하여 상단 예산 폼에 자동 입력한다."""
        site = self.cb_budget_view_site.get().strip()
        if not site:
            messagebox.showwarning("현장 미선택", "하단 '현장별 실적 조회' 영역에서 현장을 먼저 선택해주세요.")
            return

        start_ts = pd.Timestamp(self.budget_view_start.get_date())
        end_ts = pd.Timestamp(self.budget_view_end.get_date())

        if self.daily_usage_df.empty:
            messagebox.showinfo("데이터 없음", "집계할 실적 데이터가 없습니다.")
            return

        # 1. 필터링
        df = self.daily_usage_df.copy()
        df['_date_ts'] = pd.to_datetime(df['Date'], errors='coerce')
        mask = (df['_date_ts'] >= start_ts) & (df['_date_ts'] <= end_ts) & (df['Site'] == site)
        df = df[mask].copy()

        if df.empty:
            s_date = self.budget_view_start.get_date()
            e_date = self.budget_view_end.get_date()
            messagebox.showinfo("데이터 없음", f"선택한 기간 ({s_date} ~ {e_date}) 동안 '{site}' 현장의 실적이 없습니다.")
            return

        # 2. 집계 초기화
        def _f(val):
            if pd.isna(val) or str(val).lower() == 'nan': return 0.0
            try: 
                s = str(val).replace(',', '').strip()
                if not s: return 0.0
                num = float(s)
                return 0.0 if np.isnan(num) else num
            except: return 0.0

        total_net_revenue = 0.0 # 검사비
        total_travel = 0.0
        total_meal = 0.0
        
        # --- Labor Aggregation by Rank ---
        ranks = ["이사", "부장", "차장", "과장", "대리", "계장", "주임", "기사"]
        # Track unique dates for each worker to calculate active days correctly
        # Structure: {rank: {worker_name: set(date_strings)}}
        rank_labor_dates = {r: {} for r in ranks}
        
        # --- Special Work (OT) Classification ---
        special_types = ["연장근무", "야간근무", "휴일근무"]
        ot_data = {t: {'names': set(), 'hours': 0} for t in special_types}

        import re as _re
        # Updated pattern to support [Rank], Rank Name, or Name alone with mapping
        # Supports: "[부장] 주진철", "부장 주진철", "주진철"
        rank_pattern = _re.compile(r"\[?(이사|부장|차장|과장|대리|계장|주임|기사)\]?\s*(.*)")
        
        # Default mapping for core workers if rank is missing
        worker_rank_map = {
            "주진철": "부장", "우명광": "대리", "김진환": "주임", "장승대": "계장",
            "김성렬": "주임", "박광복": "부장", "주영광": "과장", "이경재": "주임",
            "황조현": "주임", "김춘호": "차장", "박원준": "과장", "이봉주": "부장",
            "김성현": "대리"
        }

        # 상세 재료비 집계를 위한 맵
        material_usage_sums = [0.0] * 10
        mat_map = {
            '세척': 0, '침투': 1, '현상': 2, '자분': 3, '흑색': 4,
            '필름': 5, '글리세': 6, '현상액': 7, '정착액': 8, '수적': 9
        }

        # 자재 마스터 정보
        mat_id_name_map = {}
        mat_id_cost_map = {}
        if not self.materials_df.empty:
            mat_id_name_map = self.materials_df.set_index('MaterialID')['품목명'].fillna('').to_dict()
            mat_id_cost_map = self.materials_df.set_index('MaterialID')['원가'].fillna(0).to_dict()

        total_mat_cost = 0.0
        rt_inspection_count = 0  # Counter for RT occurrences
        total_film_count = 0.0
        # [NEW] Use unique days instead of row count to avoid inflating days for multiple records
        total_days_count = df['_date_ts'].dt.date.nunique() if not df.empty else 0
        
        # [NEW] Track logically unique entries to avoid double-counting materials from split-row records
        processed_entry_ids = set()
        
        # [NEW] Track unique dates for vehicles and specialized equipment
        starex_dates = set()
        toptruck_dates = set()
        paut_dates = set()
        paut_manual_scanner_dates = set()
        paut_cobra_scanner_dates = set()
        mt_dates = set()
        rt_dates = set()
        # [NEW] Track unique dates for ANY equipment name found in the '장비명' column
        equip_dates_map = {}

        for _, row in df.iterrows():
            # Check weekday
            date_val = _re.sub(r'\s.*', '', str(row.get('Date', '')))
            
            try:
                dt_obj = pd.to_datetime(date_val)
                is_weekday = dt_obj.weekday() < 5 # 0-4: Mon-Fri
            except:
                is_weekday = True

            total_net_revenue += _f(row.get('검사비', 0))
            total_travel += _f(row.get('출장비', 0))
            total_meal += _f(row.get('일식', 0))

            # [NEW] Identify logically unique entries (shared save event)
            # STRATEGIC FIX: Remove 'Note' from UID because split-rows often have different notes like '(차량 추가 기록)'
            # relying only on Date + Site + EntryTime (minute level) for grouping
            raw_time = str(row.get('EntryTime', '')).strip()
            e_time = raw_time[:16] if len(raw_time) > 16 else raw_time
            e_site = str(row.get('Site', '')).strip()
            
            # UID composite - date_val is already YYYY-MM-DD
            e_uid = f"{date_val}|{e_site}|{e_time}"
            
            is_new_logical_entry = (e_uid not in processed_entry_ids)
            processed_entry_ids.add(e_uid)

            # 명칭 불일치 방지 (세척제/세척액, 침투제/침투액, 현상제/현상액 혼용 대응)
            def get_first_valid(r, keys):
                for k in keys:
                    v = _f(r.get(k, 0))
                    if v > 0: return v
                return 0.0

            # [NOTE] Specialist NDT columns are now handled in the consolidated block below to avoid double counting
            
            # --- Vehicle Tracking ---
            m_id_local = row.get('MaterialID', '')
            # [FIX] If MaterialID is a manual name string (not in master map), use it directly as the name
            m_name_local = str(mat_id_name_map.get(m_id_local, m_id_local)).upper()
            
            # [FIX] Check both Equipment Name (장비명) and Item Name (품목명) for vehicle detection
            car_info = f"{row.get('차량번호', '')} {row.get('차량비고', '')} {row.get('장비명', '')} {m_name_local}".strip()
            
            try:
                dt_obj = pd.to_datetime(date_val)
                date_key = dt_obj.date()
            except:
                date_key = date_val
                
            if is_new_logical_entry:
                if '스타렉스' in car_info:
                    starex_dates.add(date_key)
                if '탑차' in car_info:
                    toptruck_dates.add(date_key)

            # PAUT 및 특수장비 작업일수(중복일 제외) 집계
            try:
                method_val = str(row.get('검사방법', '')).upper()
                equip_name = str(row.get('장비명', '')).upper()
                
                # [FIX] Check both Equipment Name and Item Name for PAUT/MT detection
                is_paut = 'PAUT' in method_val or 'PAUT' in equip_name or 'PAUT' in m_name_local
                is_mt = 'MT' in method_val or 'MT' in equip_name or 'MT' in m_name_local or 'YOKE' in equip_name or 'YOKE' in m_name_local
                
                if is_paut:
                    paut_dates.add(date_key)
                    # Specialized Scanner Detection (Robust matching)
                    eq_clean = equip_name.replace(' ', '').upper()
                    m_clean = m_name_local.replace(' ', '').upper()
                    mid_clean = str(m_id_local).replace(' ', '').upper()
                    
                    is_manual_scanner = ('SCANNER(MANUAL)' in eq_clean or 'SCANNER(MANUAL)' in m_clean or 'SCANNER(MANUAL)' in mid_clean or
                                         ('MANUAL' in eq_clean and 'SCANNER' in eq_clean) or 
                                         ('MANUAL' in m_clean and 'SCANNER' in m_clean) or
                                         ('MANUAL' in mid_clean and 'SCANNER' in mid_clean))
                    
                    is_cobra_scanner = ('SCANNER(COBRA)' in eq_clean or 'SCANNER(COBRA)' in m_clean or 'SCANNER(COBRA)' in mid_clean or
                                        ('COBRA' in eq_clean and 'SCANNER' in eq_clean) or 
                                        ('COBRA' in m_clean and 'SCANNER' in m_clean) or
                                        ('COBRA' in mid_clean and 'SCANNER' in mid_clean))

                    if is_manual_scanner:
                        paut_manual_scanner_dates.add(date_key)
                    if is_cobra_scanner:
                        paut_cobra_scanner_dates.add(date_key)
                        
                if is_mt:
                    mt_dates.add(date_key)
                if 'RT' in method_val or 'RT' in equip_name or 'RT' in m_name_local:
                    rt_dates.add(date_key)
            except Exception:
                pass

            # --- Labor & OT Parsing ---
            shift = str(row.get('Shift', '주간')).strip()
            for i in range(1, 11):
                suffix = '' if i == 1 else str(i)
                raw_worker = str(row.get(f'User{suffix}', '')).strip()
                if not raw_worker or raw_worker.lower() == 'nan':
                    continue
                
                # Rank Aggregation
                match = rank_pattern.search(raw_worker)
                worker_name_only = raw_worker
                rank = None
                
                if match:
                    # Case 1: Rank identified via pattern (e.g., "[부장] 주진철" or "부장 주진철")
                    if match.group(1) in ranks:
                        rank = match.group(1)
                        worker_name_only = match.group(2).strip()
                        if not worker_name_only: # Only rank was present
                             worker_name_only = rank
                
                # Case 2: Fallback to mapping if rank not identified from the string
                if not rank:
                    # Clean potential shift prefixes if they exist (though usually cleaned in records)
                    clean_name = _re.sub(r'\(.*?\)', '', raw_worker).strip()
                    rank = worker_rank_map.get(clean_name)
                    worker_name_only = clean_name
                
                if rank and rank in rank_labor_dates:
                    if worker_name_only not in rank_labor_dates[rank]:
                        rank_labor_dates[rank][worker_name_only] = set()
                    try:
                        # Normalize date key for uniqueness
                        dt_obj = pd.to_datetime(date_val)
                        d_key = dt_obj.date()
                    except:
                        d_key = date_val
                    rank_labor_dates[rank][worker_name_only].add(d_key)
                
                # [FIX] WorkTime 컬럼에서 분류(주간/야간/휴일)와 시간 파싱
                # 형식: "(야간) 4h" or "(주야간) 2.5h" 또는 "8h" (구형식)
                wt_col = f'WorkTime{suffix}'
                wt_val = str(row.get(wt_col, '')).strip()
                
                # WorkTime에서 분류 파싱
                wt_shift = shift  # 기본값: 레코드의 Shift 컬럼
                wt_hours = 0.0
                if wt_val and wt_val.lower() != 'nan':
                    # Parse shift type if present, e.g. "(야간)", "(연장)"
                    shift_match = _re.search(r'\(?(주간|야간|주야간|휴일|연장)\)?', wt_val)
                    if shift_match:
                        wt_shift = shift_match.group(1)
                    
                    # Compute duration calculation
                    marker_pattern = MARKER_PATTERN
                    clean_wt = marker_pattern.sub('', wt_val).replace('(연장)', '').replace('(주야간)', '').replace('(야간)', '').replace('(휴일)', '').replace('(주간)', '').strip()
                    sep_match = _re.search(r'[:\d]\s*([~-])\s*[:\d]', clean_wt)
                    
                    if sep_match:
                        try:
                            sep = sep_match.group(1)
                            st_str, en_str = clean_wt.split(sep)
                            sh, sm = map(int, st_str.split(':'))
                            eh, em = map(int, en_str.split(':'))
                            start_f = sh + sm/60.0
                            end_f = eh + em/60.0
                            if end_f < start_f: end_f += 24
                            wt_hours = end_f - start_f
                        except:
                            pass
                    else:
                        # Fallback: Just a number like "8h" or "8"
                        h_match = _re.search(r'([\d.]+)', clean_wt)
                        if h_match:
                            try: wt_hours = float(h_match.group(1))
                            except: pass
                
                # [FIX] 분류별 OT 상세 집계
                if wt_hours > 0:
                    # 명시적으로 '휴일'이라고 쓰여있거나, 주말인데 기본 교대근무가 아닌 특근성 출근인 경우 (이번엔 단순히 텍스트 명시에 더 의존)
                    if '휴일' in wt_shift:
                        ot_data['휴일근무']['hours'] += wt_hours
                        ot_data['휴일근무']['names'].add(worker_name_only)
                    else:
                        # 평일/주말 공통: 연장이나 야간 명시가 있거나 시간대 기반 계산
                        if sep_match:
                            try:
                                sep = sep_match.group(1)
                                st_str, en_str = clean_wt.split(sep)
                                sh, sm = map(int, st_str.split(':'))
                                eh, em = map(int, en_str.split(':'))
                                start_f = sh + sm/60.0
                                end_f = eh + em/60.0
                                if end_f < start_f: end_f += 24
                                
                                # 현장 기본 퇴근 시간 (17시 기준 혹은 18시 등)
                                # 전체 시간이 주어질 때 명확히 구분 (기본근무: ~18, 연장근무: 18~22, 야간근무: 22~)
                                if end_f > 18:
                                    ot_start = max(start_f, 18.0)
                                    # 18~22: 연장근무
                                    h_over = max(0, min(end_f, 22.0) - ot_start)
                                    # 22~24 : 야간근무
                                    h_night = max(0, min(end_f, 24.0) - max(ot_start, 22.0))
                                    # 24~ : 심야/익일 근무
                                    h_dawn = max(0, end_f - max(ot_start, 24.0))
                                    
                                    if h_over > 0:
                                        ot_data['연장근무']['hours'] += h_over
                                        ot_data['연장근무']['names'].add(worker_name_only)
                                    if h_night > 0:
                                        ot_data['야간근무']['hours'] += h_night
                                        ot_data['야간근무']['names'].add(worker_name_only)
                                    if h_dawn > 0:
                                        try:
                                            # 날짜 확인 (금요일 밤 -> 토요일 새벽인지)
                                            entry_date = pd.to_datetime(row.get('Date'))
                                            if entry_date.weekday() == 4: # 4 = 금요일
                                                ot_data['휴일근무']['hours'] += h_dawn
                                                ot_data['휴일근무']['names'].add(worker_name_only)
                                            else:
                                                ot_data['야간근무']['hours'] += h_dawn
                                                ot_data['야간근무']['names'].add(worker_name_only)
                                        except:
                                            ot_data['야간근무']['hours'] += h_dawn
                                            ot_data['야간근무']['names'].add(worker_name_only)
                            except: pass
                        else:
                            # 시간 범위가 주어지지 않은 숫자(예: 4h, 10h) 처리 로직
                            # 명시적으로 (연장) 태그가 있는 경우
                            if '연장' in wt_shift:
                                ot_data['연장근무']['hours'] += wt_hours
                                ot_data['연장근무']['names'].add(worker_name_only)
                            # 명시적 태그가 없으나 8시간을 초과하는 경우
                            elif wt_hours > 8:
                                h_extra = wt_hours - 8
                                if '야간' in wt_shift or '주야간' in wt_shift:
                                    ot_data['야간근무']['hours'] += h_extra
                                    ot_data['야간근무']['names'].add(worker_name_only)
                                else:
                                    ot_data['연장근무']['hours'] += h_extra
                                    ot_data['연장근무']['names'].add(worker_name_only)


            # --- 자재 사용량 집계 (NDT + 필름) ---
            if is_new_logical_entry:
                m_id = row.get('MaterialID', '')
                m_name = str(mat_id_name_map.get(m_id, '')).strip()
                test_method = str(row.get('검사방법', '')).upper()
                usage = _f(row.get('Usage', 0))
                
                if 'RT' in test_method:
                    rt_inspection_count += 1
                    film_qty = get_first_valid(row, ['필름매수', 'FilmCount', 'Usage', '검사량'])
                    total_film_count += film_qty
                    material_usage_sums[5] += film_qty  # Index 5 = 필름
                    total_mat_cost += film_qty * float(mat_id_cost_map.get(m_id, 0))
                else:
                    # UT/PT/MT 검사: NDT 화학약품 등의 일반 자재 사용
                    # Specialist columns take priority over the generic 'Usage' field
                    ndt_vals = [
                        get_first_valid(row, ['NDT_세척제', '세척제', 'NDT_세척액', '세척액']),
                        get_first_valid(row, ['NDT_침투제', '침투제', 'NDT_침투액', '침투액']),
                        get_first_valid(row, ['NDT_현상제', '현상제', 'NDT_현상액', '현상액']),
                        get_first_valid(row, ['NDT_흑색자분', 'NDT_자분', '흑색자분', 'NDT_페인트']),
                        get_first_valid(row, ['NDT_백색페인트', '백색페인트']) # Placeholder mapping adjustment
                    ]
                    
                    has_specialist_data = any(v > 0 for v in ndt_vals)
                    
                    if has_specialist_data:
                        for idx, v in enumerate(ndt_vals):
                            if idx < 5: material_usage_sums[idx] += v
                        # Cost calculation still needs a MaterialID context, fallback to usage row if available
                        total_mat_cost += usage * float(mat_id_cost_map.get(m_id, 0))
                    elif usage > 0:
                        # Fallback to Usage field matching if specialist columns are empty
                        # 세척제: 청소/세척 용도 → Index 0
                        if '세척제' in m_name or '세척' in m_name or '청소' in m_name:
                            material_usage_sums[0] += usage
                        # 침투제: 침투액 → Index 1
                        elif '침투제' in m_name or '침투액' in m_name:
                            material_usage_sums[1] += usage
                        # 현상제: 현상용 시약 → Index 2
                        elif '현상제' in m_name or '현상액' in m_name:
                            material_usage_sums[2] += usage
                        # 자분: 자기 자분 → Index 3
                        elif '자분' in m_name and '흑색' not in m_name:
                            material_usage_sums[3] += usage
                        # 흑색자분 → Index 4
                        elif '흑색자분' in m_name or ('흑색' in m_name and '자분' in m_name):
                            material_usage_sums[4] += usage
                        # 글리세린 → Index 6
                        elif '글리세' in m_name or '글리세린' in m_name:
                            material_usage_sums[6] += usage
                        # 정착액 → Index 8
                        elif '정착액' in m_name or '정착' in m_name:
                            material_usage_sums[8] += usage
                        # 수적 → Index 9
                        elif '수적' in m_name:
                            material_usage_sums[9] += usage
                        # 기타 PT/MT 화학약품이면 자동 매핑
                        elif any(k in m_name.upper() for k in ['PENETRANT', 'DEVELOPER', 'CLEANER', 'NABAKEM', 'PT', 'MT']):
                            matched = False
                            for keyword, idx in mat_map.items():
                                if keyword.lower() in m_name.lower():
                                    material_usage_sums[idx] += usage
                                    matched = True
                                    break
                            if not matched:
                                material_usage_sums[0] += usage
                        total_mat_cost += usage * float(mat_id_cost_map.get(m_id, 0))

        # 3. 폼에 입력 및 [5. 영업이익] '실행(실적)' 줄에 표시 결과 연동
        # Update Special Unit Prices if LaborCostDetailWidget is available to compute correct revenue
        ot_revenue = 0.0
        if hasattr(self, 'labor_detail_widget'):
            ot_revenue = sum(ot_data[t]['hours'] * _f(self.labor_detail_widget.entries[t]['unit_price'].get()) 
                            for t in special_types if t in self.labor_detail_widget.entries)

        self.cb_budget_site.set(site)
        
        # [NEW] 도급액(Revenue) 칸에 실적 검사비 총액 자동 기입
        self.ent_budget_revenue.delete(0, tk.END)
        self.ent_budget_revenue.insert(0, f"{total_net_revenue:,.0f}")

        # [NEW] 실적 데이터의 검사비 총액을 '검사단가' 칸에 기본 입력
        if hasattr(self, 'ent_budget_unit_price'):
            self.ent_budget_unit_price.delete(0, tk.END)
            self.ent_budget_unit_price.insert(0, f"{total_net_revenue:,.0f}")

        self.ent_budget_material.delete(0, tk.END)
        self.ent_budget_material.insert(0, f"{np.nan_to_num(total_mat_cost):,.0f}")

        self.ent_budget_expense.delete(0, tk.END)
        self.ent_budget_expense.insert(0, f"{np.nan_to_num(total_travel + total_meal):,.0f}")


        # 4. 인건비 상세 위젯 업데이트
        if hasattr(self, 'labor_detail_widget'):
            labor_data = self.labor_detail_widget.get_data()
            # Regular
            for rank in ranks:
                if rank in labor_data:
                    # Number of unique workers who logged time in this rank
                    u_count = len(rank_labor_dates[rank])
                    # Total number of unique working days summed across all workers in this rank
                    t_days = sum(len(dates) for dates in rank_labor_dates[rank].values())
                    if u_count > 0:
                        avg_d = t_days / u_count
                        labor_data[rank]['personnel'] = f"{u_count:g}"
                        labor_data[rank]['period'] = f"{avg_d:g}"
                    else:
                        labor_data[rank]['personnel'] = ""
                        labor_data[rank]['period'] = ""
            # Special (OT)
            for stype in special_types:
                if stype in labor_data:
                    u_count = len(ot_data[stype]['names'])
                    t_hours = ot_data[stype]['hours']
                    if u_count > 0:
                        avg_h = t_hours / u_count
                        labor_data[stype]['personnel'] = f"{u_count:g}"
                        labor_data[stype]['period'] = f"{avg_h:g}"
                    else:
                        labor_data[stype]['personnel'] = ""
                        labor_data[stype]['period'] = ""
            self.labor_detail_widget.set_data(labor_data)

        # 5. 재료비 상세 위젯 업데이트
        if hasattr(self, 'material_detail_widget'):
            current_mat_data = self.material_detail_widget.get_data()

            # [FIX] 필름현상액/정착액 자동 계산
            # 기준: 필름 500매 -> 각 2개, 1000매 -> 각 4개
            # 즉 250매당 1개
            film_chem_qty = (total_film_count / 250.0) if total_film_count > 0 else 0.0
            # 수적방지액(200mL): 필름 500매당 1개
            anti_drop_qty = (total_film_count / 500.0) if total_film_count > 0 else 0.0
            if len(material_usage_sums) > 7:
                material_usage_sums[7] = film_chem_qty  # 필름현상액 3L
            if len(material_usage_sums) > 8:
                material_usage_sums[8] = film_chem_qty  # 정착액 3L
            if len(material_usage_sums) > 9:
                material_usage_sums[9] = anti_drop_qty  # 수적방지액 200mL

            for i, qty in enumerate(material_usage_sums):
                if i < len(current_mat_data):
                    current_mat_data[i]['qty'] = f"{qty:g}" if qty > 0 else ""
            self.material_detail_widget.set_data(current_mat_data)
        
        # [NEW] Immediate KPI update after data fill
        self._update_budget_kpis()
            
        # 6. 경비 상세 위젯 업데이트 (차량유지비, 소모품비, 복리후생비, Se-175 수량, 감가상각비 차량 사용일수 자동 할당)
        if hasattr(self, 'expense_detail_widget'):
            self.expense_detail_widget.calculate_all()
            
            # [DEFINITIVE FIX] Safely sync widget totals back to the main form entries
            if hasattr(self, 'labor_detail_widget') and hasattr(self, 'ent_budget_labor'):
                lab_total = 0.0
                try:
                    # Try method first
                    if hasattr(self.labor_detail_widget, 'get_total_cost'):
                        lab_total = self.labor_detail_widget.get_total_cost()
                    else:
                        # Direct label fallback
                        raw_text = self.labor_detail_widget.lbl_grand_total.cget('text')
                        val = "".join(c for c in raw_text if c.isdigit() or c == '.')
                        lab_total = float(val or 0)
                except: pass
                
                self.ent_budget_labor.delete(0, tk.END)
                self.ent_budget_labor.insert(0, f"{lab_total:,.0f}")

            if hasattr(self, 'expense_detail_widget') and hasattr(self, 'ent_budget_expense'):
                exp_total = 0.0
                try:
                    if hasattr(self.expense_detail_widget, 'get_total_cost'):
                        exp_total = self.expense_detail_widget.get_total_cost()
                    else:
                        raw_text = self.expense_detail_widget.lbl_exp_total.cget('text')
                        val = "".join(c for c in raw_text if c.isdigit() or c == '.')
                        exp_total = float(val or 0)
                except: pass
                
                self.ent_budget_expense.delete(0, tk.END)
                self.ent_budget_expense.insert(0, f"{exp_total:,.0f}")

            # [STABILITY] Final KPI and Profit sync refresh
            self._update_budget_kpis()
            self.root.update_idletasks()
            current_exp_data = self.expense_detail_widget.get_data()

            # [FIX] 현장경비 수량(qty) 자동 할당 (기본값 무관하게 실적 기준 덮어쓰기)
            for row_data in current_exp_data.get('site_expense', []):
                cat = str(row_data.get('cat', '')).upper()
                if any(k in cat for k in ['차량유지비', '소모품비', '복리후생비']):
                    row_data['qty'] = f"{total_days_count:g}" if total_days_count > 0 else ""
                elif 'SE-175' in cat or 'SE175' in cat:
                    # [REVERT] Count unique work days instead of total inspection quantity
                    row_data['qty'] = f"{len(rt_dates):g}" if len(rt_dates) > 0 else ""
                    
            for row_data in current_exp_data.get('depreciation', []):
                item = row_data.get('item', '')
                if '스타렉스' in item:
                    row_data['days'] = f"{len(starex_dates):g}" if len(starex_dates) > 0 else ""
                elif '탑차' in item:
                    row_data['days'] = f"{len(toptruck_dates):g}" if len(toptruck_dates) > 0 else ""
                item_upper_clean = str(item).upper().replace(' ', '')
                if 'PAUTSCANNER(MANUAL)' in item_upper_clean:
                    row_data['days'] = f"{len(paut_manual_scanner_dates):g}" if len(paut_manual_scanner_dates) > 0 else ""
                elif 'PAUTSCANNER(COBRA)' in item_upper_clean:
                    row_data['days'] = f"{len(paut_cobra_scanner_dates):g}" if len(paut_cobra_scanner_dates) > 0 else ""
                elif 'PAUT' in item_upper_clean:
                    row_data['days'] = f"{len(paut_dates):g}" if len(paut_dates) > 0 else ""
                elif 'YOKE' in str(item).upper() or ('MT' in str(item).upper() and 'PAUT' not in str(item).upper()):
                    row_data['days'] = f"{len(mt_dates):g}" if len(mt_dates) > 0 else ""
                else:
                    # [NEW] Check for generic equipment name matching (case-insensitive)
                    item_upper = str(item).upper().strip()
                    matched_days = 0
                    for e_name, e_dates in equip_dates_map.items():
                        if e_name.upper().strip() in item_upper or item_upper in e_name.upper().strip():
                            matched_days = max(matched_days, len(e_dates))
                    
                    if matched_days > 0:
                        row_data['days'] = f"{matched_days:g}"
                    
            self.expense_detail_widget.set_data(current_exp_data)

        # [NEW] Calculate profit/margin for the summary message
        try:
            rev = float(self.ent_budget_unit_price.get().replace(',', '') or 0)
            lab = float(self.ent_budget_labor.get().replace(',', '') or 0)
            mat = float(self.ent_budget_material.get().replace(',', '') or 0)
            exp = float(self.ent_budget_expense.get().replace(',', '') or 0)
            out = float(self.ent_budget_outsource.get().replace(',', '') or 0)
            cost = lab + mat + exp + out
            profit = rev - cost
            margin = (profit / rev * 100) if rev > 0 else 0.0
            profit_str = f"\n\n▶ 집계 결과\n- 도급액(검사비): {rev:,.0f}원\n- 실행원가: {cost:,.0f}원\n- 영업이익: {profit:,.0f}원 ({margin:.1f}%)"
        except:
            profit_str = ""

        messagebox.showinfo("조회 완료", 
                            f"'{site}' 현장의 실적 데이터를 집계하여 [5. 영업이익] 탭에 표시하였습니다.\n"
                            f"(기간: {start_ts.strftime('%Y-%m-%d')} ~ {end_ts.strftime('%Y-%m-%d')}){profit_str}")

    def save_budget_entry(self):
        """Save or update budget entry"""
        site = self.cb_budget_site.get().strip()
        if not site:
            messagebox.showwarning("입력 오류", "현장명을 입력해주세요.")
            return
            
        try:
            rev = float(str(self.ent_budget_revenue.get()).replace(',', '') or 0)
            unit_price_val = str(getattr(self, 'ent_budget_unit_price', tk.Entry()).get() if hasattr(self, 'ent_budget_unit_price') else '').replace(',', '').strip()
            unit_price = float(unit_price_val) if unit_price_val else 0.0
            lab = float(str(self.ent_budget_labor.get()).replace(',', '') or 0)
            mat = float(str(self.ent_budget_material.get()).replace(',', '') or 0)
            exp = float(str(self.ent_budget_expense.get()).replace(',', '') or 0)
            out = float(str(self.ent_budget_outsource.get()).replace(',', '') or 0)
            note = self.ent_budget_note.get().strip()
            profit = rev - (lab + mat + exp + out)
        except ValueError:
            messagebox.showerror("입력 오류", "금액은 숫자여야 합니다.")
            return

        new_data = {
            'Site': site,
            'Revenue': rev,
            'UnitPrice': unit_price,
            'LaborCost': lab,
            'MaterialCost': mat,
            'Expense': exp,
            'OutsourceCost': out,
            'Profit': profit,
            'Note': note,
            'LaborDetail': json.dumps(self.labor_detail_widget.get_data()),
            'MaterialDetail': json.dumps(self.material_detail_widget.get_data()),
            'ExpenseDetail': json.dumps(self.expense_detail_widget.get_data())
        }

        # [NEW] Robust Sanitization: Ensure NO NaN or empty strings ever reach the float64 columns
        # This fixes TypeError: Invalid value '' for dtype 'float64' permanently
        numeric_keys = {'Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit'}
        new_data_sanitized = {}
        for k, v in new_data.items():
            if k in numeric_keys:
                try:
                    # Coerce everything to float for numeric columns
                    if pd.isna(v) or str(v).strip().lower() in ('nan', 'none', ''):
                        new_data_sanitized[k] = 0.0
                    else:
                        new_data_sanitized[k] = float(v)
                except:
                    new_data_sanitized[k] = 0.0
            else:
                # String columns (Note, Detail JSONs, Site)
                if pd.isna(v) or str(v).lower() == 'nan':
                    new_data_sanitized[k] = ""
                else:
                    new_data_sanitized[k] = str(v)


        if not self.budget_df.empty and site in self.budget_df['Site'].values:
            # [FIX] Use explicit .loc[index, key] to ensure correct row update
            idx = self.budget_df[self.budget_df['Site'] == site].index[0]
            for key, val in new_data_sanitized.items():
                self.budget_df.loc[idx, key] = val
        else:
            # Create a properly typed DataFrame for concatenation to avoid dtype clashes
            row_df = pd.DataFrame([new_data_sanitized])
            self.budget_df = pd.concat([self.budget_df, row_df], ignore_index=True)
            
        if self.save_data():
            messagebox.showinfo("성공", f"'{site}' 현장의 예산 정보가 저장/수정되었습니다.")
            
            # [NEW] Refresh UI to show the updated data in comparison rows immediately
            self.update_budget_view()
            
            if 'budget' in self.detached_windows: self._refresh_detached_budget()
            # [NEW] 저장 후 현장 목록 갱신 (새로 기입한 현장명 반영)
            self.refresh_inquiry_filters()
            # self.clear_budget_form() # [UX] Don't clear immediately so user can verify

    def _toggle_hidden_site(self, site, hide=True):
        """현장명을 드롭다운 목록에서 숨기거나 다시 표시한다."""
        if not hasattr(self, 'hidden_sites'):
            self.hidden_sites = []
        if hide:
            if site not in self.hidden_sites:
                self.hidden_sites.append(site)
            messagebox.showinfo("숨기기 완료", f"'{site}' 현장이 목록에서 숨겨졌습니다.\n다시 표시하려면 현장명 입력칸을 우클릭하세요.")
        else:
            if site in self.hidden_sites:
                self.hidden_sites.remove(site)
            messagebox.showinfo("표시 완료", f"'{site}' 현장이 목록에 다시 표시됩니다.")
        # config 저장 및 목록 갱신
        self.save_tab_config()
        self.refresh_inquiry_filters()

    def _show_all_hidden_sites(self):
        """숨긴 현장명을 모두 다시 표시한다."""
        if not getattr(self, 'hidden_sites', []):
            messagebox.showinfo("알림", "숨겨진 현장이 없습니다.")
            return
        count = len(self.hidden_sites)
        names = '\n'.join(f'  • {s}' for s in self.hidden_sites)
        confirmed = messagebox.askyesno(
            "모두 표시",
            f"현재 숨겨진 현장 {count}개:\n{names}\n\n모두 다시 표시하시겠습니까?"
        )
        if confirmed:
            self.hidden_sites.clear()
            self.save_tab_config()
            self.refresh_inquiry_filters()
            messagebox.showinfo("완료", "모든 현장이 다시 표시됩니다.")

    def delete_budget_entry(self):
        """Delete selected budget entry"""
        site = self.cb_budget_site.get().strip()
        if not site:
            messagebox.showwarning("삭제 오류", "삭제할 현장명을 입력해주세요.")
            return
            
        if messagebox.askyesno("삭제 확인", f"'{site}' 현장의 예산 정보를 삭제하시겠습니까?"):
            self.budget_df = self.budget_df[self.budget_df['Site'] != site]
            if self.save_data():
                messagebox.showinfo("성공", "삭제되었습니다.")
                self.update_budget_view()
                if 'budget' in self.detached_windows: self._refresh_detached_budget()
                # [NEW] 삭제 후 현장 목록 갱신
                self.refresh_inquiry_filters()
                self.clear_budget_form()


    def clear_budget_form(self):
        """Reset budget form fields while maintaining site selection context"""
        # self.cb_budget_site.set('')  <- Removed to preserve 'appearance' as per user request
        self.ent_budget_revenue.delete(0, tk.END)
        if hasattr(self, 'ent_budget_unit_price'):
            self.ent_budget_unit_price.delete(0, tk.END)
        self.ent_budget_labor.delete(0, tk.END)
        self.ent_budget_material.delete(0, tk.END)
        self.ent_budget_expense.delete(0, tk.END)
        self.ent_budget_outsource.delete(0, tk.END)
        if hasattr(self, 'ent_budget_profit'):
            self.ent_budget_profit.delete(0, tk.END)
        self.ent_budget_note.delete(0, tk.END)
        # [NEW] Reset detail widgets
        if hasattr(self, 'labor_detail_widget'): self.labor_detail_widget.reset()
        if hasattr(self, 'material_detail_widget'):
            self.material_detail_widget.reset()
        if hasattr(self, 'expense_detail_widget'):
            self.expense_detail_widget.reset()
            
        # [STABILITY] Reset KPI summary
        self._update_budget_kpis()

    def on_budget_tree_select(self, event):
        """No-op - 하단 목록 제거됨"""
        pass

    def _ensure_sash_visible(self):
        """No-op as details panel is removed"""
        pass
    
    def _ensure_daily_usage_sash_visibility(self):
        """Ensure sash position is visible and properly sized using ratio"""
        try:
            if not hasattr(self, 'daily_usage_paned'):
                return
                
            # [STABILITY] Skip if we are still in the initial loading/restoration phase
            # load_tab_config will handle the initial placement.
            if not getattr(self, 'is_ready', False):
                return

            # Force multiple updates to get correct dimensions
            for _ in range(3):
                self.daily_usage_paned.update_idletasks()
            
            # Get current paned window dimensions
            total_h = self.daily_usage_paned.winfo_height()
            
            if total_h <= 0:
                print("PanedWindow height is 0, retrying...")
                self.root.after(100, self._ensure_daily_usage_sash_visibility)
                return
            
            # If locked, prioritize absolute position if valid
            if hasattr(self, 'daily_usage_sash_locked') and self.daily_usage_sash_locked:
                if hasattr(self, 'tab_config') and 'daily_usage_sash_pos' in self.tab_config:
                    target_pos = int(self.tab_config['daily_usage_sash_pos'])
                    # Ensure within bounds
                    if 50 < target_pos < total_h - 50:
                        self.daily_usage_paned.sashpos(0, target_pos)
                        print(f"Restored locked absolute position: {target_pos}")
                        return

            # Try to restore from saved ratio first
            # [UX FIX] 작업자 박스 하단이 가려지지 않도록 상단 영역 기본 비율 상향
            target_ratio = 0.45  # Default ratio
            if hasattr(self, 'tab_config') and 'daily_usage_sash_ratio' in self.tab_config:
                target_ratio = self.tab_config['daily_usage_sash_ratio']
            
            # Calculate target position
            target_pos = int(total_h * target_ratio)
            
            # Apply relaxed bounds + practical minimum top area (Reduced from 460 to 100 to allow user custom sizing)
            min_pos = max(int(total_h * 0.1), 100)
            max_pos = int(total_h * 0.95)
            
            if target_pos < min_pos:
                target_pos = min_pos
            elif target_pos > max_pos:
                target_pos = max_pos
            
            # Set sash position
            self.daily_usage_paned.sashpos(0, target_pos)
            
            # Verify position was set correctly
            actual_pos = self.daily_usage_paned.sashpos(0)
            actual_ratio = actual_pos / total_h if total_h > 0 else target_ratio
            
            # Update config
            if hasattr(self, 'is_ready') and self.is_ready:
                if not hasattr(self, 'tab_config'):
                    self.tab_config = {}
                self.tab_config['daily_usage_sash_ratio'] = actual_ratio
                self.tab_config['daily_usage_sash_pos'] = actual_pos
            
            # Update canvas scroll region
            self.root.after(100, self._ensure_canvas_scroll_region)
            
        except Exception as e:
            print(f"Error ensuring sash visibility: {e}")
            self.root.after(500, self._ensure_daily_usage_sash_visibility)
    
    
    
    def update_resolution_display(self):
        """Update the resolution display"""
        try:
            if hasattr(self, 'resolution_label'):
                # Get window dimensions
                window_width = self.root.winfo_width()
                window_height = self.root.winfo_height()
                
                # Get screen dimensions
                screen_width = self.root.winfo_screenwidth()
                screen_height = self.root.winfo_screenheight()
                
                # Get daily usage paned window dimensions
                if hasattr(self, 'daily_usage_paned'):
                    self.daily_usage_paned.update_idletasks()
                    pane_height = self.daily_usage_paned.winfo_height()
                    pane_width = self.daily_usage_paned.winfo_width()
                    
                    # Get current sash position and ratio
                    try:
                        sash_pos = self.daily_usage_paned.sashpos(0)
                        ratio = (sash_pos / pane_height * 100) if pane_height > 0 else 0
                        resolution_text = f"창: {window_width}x{window_height} | 화면: {screen_width}x{screen_height} | 패널: {pane_width}x{pane_height} | 경계: {ratio:.1f}%"
                    except:
                        resolution_text = f"창: {window_width}x{window_height} | 화면: {screen_width}x{screen_height} | 패널: {pane_width}x{pane_height}"
                else:
                    resolution_text = f"창: {window_width}x{window_height} | 화면: {screen_width}x{screen_height}"
                
                self.resolution_label.config(text=resolution_text)
                
                # Schedule next update
                self.root.after(500, self.update_resolution_display)
        except Exception as e:
            print(f"Error updating resolution display: {e}")
            # Still schedule next update even if there's an error
            if hasattr(self, 'resolution_label'):
                self.root.after(1000, self.update_resolution_display)
    
    def _restore_locked_position(self):
        """Restore sash to locked position"""
        try:
            if not hasattr(self, 'daily_usage_paned'): return
            
            if hasattr(self, 'tab_config'):
                total_h = self.daily_usage_paned.winfo_height()
                if total_h <= 0: return
                
                # Prioritize absolute position when locked
                if 'daily_usage_sash_pos' in self.tab_config:
                    pos = int(self.tab_config['daily_usage_sash_pos'])
                    # Only apply if it doesn't hide the bottom area completely
                    if 50 < pos < total_h - 50:
                        self.daily_usage_paned.sashpos(0, pos)
                        return
                
                # Fallback to ratio
                if 'daily_usage_sash_ratio' in self.tab_config:
                    ratio = self.tab_config['daily_usage_sash_ratio']
                    locked_pos = int(total_h * ratio)
                    self.daily_usage_paned.sashpos(0, locked_pos)
        except Exception as e:
            print(f"Error restoring locked position: {e}")
    
    def _on_daily_usage_resize(self, event):
        """Handle window resize to maintain sash ratio or absolute position if locked"""
        try:
            if not hasattr(self, 'daily_usage_paned'): return
            
            # If locked, maintain absolute position from top
            if hasattr(self, 'daily_usage_sash_locked') and self.daily_usage_sash_locked:
                self._restore_locked_position()
                return

            # Otherwise maintain ratio
            if hasattr(self, 'tab_config') and 'daily_usage_sash_ratio' in self.tab_config:
                ratio = self.tab_config['daily_usage_sash_ratio']
                total_h = self.daily_usage_paned.winfo_height()
                
                if total_h > 200:
                    new_pos = int(total_h * ratio)
                    min_pos, max_pos = 50, total_h - 50
                    new_pos = max(min_pos, min(new_pos, max_pos))
                    
                self.daily_usage_paned.sashpos(0, new_pos)
        except Exception as e:
            print(f"Error handling resize: {e}")
            # Fallback if something went wrong during resize

    def toggle_sash_lock(self):
        """Toggle sash lock state with auto-fit logic"""
        try:
            self.daily_usage_sash_locked = not self.daily_usage_sash_locked
            
            if self.daily_usage_sash_locked:
                # [NEW] Auto-fit sash to content height when locking
                self.root.update_idletasks()
                
                # Calculate required height of the top area content
                content_h = 0
                if hasattr(self, 'entry_inner_frame'):
                    # Use grid_bbox of the inner frame
                    bbox = self.entry_inner_frame.grid_bbox()
                    content_h = bbox[1] + bbox[3] + 5 # Minimized padding
                
                total_height = self.daily_usage_paned.winfo_height()
                
                # If content is smaller than total height, fit to content
                # Otherwise, use current sash position
                if 460 < content_h < total_height - 100:
                    sash_pos = content_h
                else:
                    sash_pos = self.daily_usage_paned.sashpos(0)
                
                # Apply sash position
                self.daily_usage_paned.sashpos(0, sash_pos)
                
                if not hasattr(self, 'tab_config'):
                    self.tab_config = {}
                
                # Calculate and save ratio and absolute position
                ratio = sash_pos / total_height if total_height > 0 else 0.5
                self.tab_config['daily_usage_sash_ratio'] = ratio
                self.tab_config['daily_usage_sash_pos'] = sash_pos
                self.tab_config['daily_usage_sash_locked'] = True
                
                # Save configuration immediately
                self.save_tab_config(force=True)
                
                if hasattr(self, 'btn_sash_lock'):
                    self.btn_sash_lock.config(text="🔒 경계 고정됨")
                    self.btn_sash_lock.configure(style="SashLock.TButton")
                    self.style.configure("SashLock.TButton", foreground="red")
                
                self._start_sash_monitor()
                print(f"Daily usage sash position LOCKED at height: {sash_pos}")
            else:
                # Unlock
                if hasattr(self, 'btn_sash_lock'):
                    self.btn_sash_lock.config(text="🔓 경계 자유")
                    self.btn_sash_lock.configure(style="TButton")
                
                self._stop_sash_monitor()
                if hasattr(self, 'tab_config'):
                    self.tab_config['daily_usage_sash_locked'] = False
                    self.save_tab_config(force=True)
                print("Daily usage sash position UNLOCKED")
            
            # [NEW] Refresh scroll region to ensure it stops correctly at the new boundary
            self._ensure_canvas_scroll_region()
        except Exception as e:
            print(f"Error toggling sash lock: {e}")
    
    def _start_sash_monitor(self):
        """Start periodic monitoring of sash position"""
        if hasattr(self, '_sash_monitor_job'):
            self.root.after_cancel(self._sash_monitor_job)
        
        def check_sash():
            if hasattr(self, 'daily_usage_sash_locked') and self.daily_usage_sash_locked:
                self._restore_locked_position()
                # Schedule next check
                self._sash_monitor_job = self.root.after(200, check_sash)
        
        # Start monitoring
        self._sash_monitor_job = self.root.after(200, check_sash)
        print("Started sash position monitoring")
    
    def _stop_sash_monitor(self):
        """Stop periodic monitoring of sash position"""
        if hasattr(self, '_sash_monitor_job'):
            self.root.after_cancel(self._sash_monitor_job)
            delattr(self, '_sash_monitor_job')
            print("Stopped sash position monitoring")
    
    def _on_main_window_resize(self, event):
        """Handle main window resize to maintain all sash ratios"""
        try:
            # Only process for actual window resize, not widget events
            if event.widget == self.root:
                # Check if daily usage tab exists and has saved ratio
                if hasattr(self, 'daily_usage_paned') and hasattr(self, 'tab_config') and 'daily_usage_sash_ratio' in self.tab_config:
                    self.root.after(100, self._ensure_daily_usage_sash_visibility)
                
                # Check if inout tab exists and has saved ratio
                if hasattr(self, 'inout_paned') and hasattr(self, 'tab_config') and 'inout_sash_ratio' in self.tab_config:
                    self.root.after(100, self._ensure_inout_sash_visibility)
                    
        except Exception as e:
            print(f"Error handling main window resize: {e}")
    
    
    def show_error_dialog(self, title, message):
        """Show a custom error dialog with draggable text"""
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("600x400")
        dialog.resizable(True, True)
        
        # Make dialog modal
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (600 // 2)
        y = (dialog.winfo_screenheight() // 2) - (400 // 2)
        dialog.geometry(f"600x400+{x}+{y}")
        
        # Main frame
        main_frame = ttk.Frame(dialog)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Title label
        title_label = ttk.Label(main_frame, text=title, font=('Malgun Gothic', 12, 'bold'))
        title_label.pack(pady=(0, 10))
        
        # Text widget with scrollbar for error message
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill='both', expand=True)
        
        text_widget = tk.Text(text_frame, wrap='word', font=('Malgun Gothic', 10))
        scrollbar = ttk.Scrollbar(text_frame, orient='vertical', command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Insert error message
        text_widget.insert('1.0', message)
        text_widget.configure(state='normal')  # Allow selection and copying
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill='x', pady=(10, 0))
        
        # Copy button
        def copy_text():
            dialog.clipboard_clear()
            selected_text = text_widget.get('sel.first', 'sel.last') if text_widget.tag_ranges('sel') else text_widget.get('1.0', 'end')
            dialog.clipboard_append(selected_text)
        
        copy_btn = ttk.Button(button_frame, text="복사하기", command=copy_text)
        copy_btn.pack(side='left', padx=5)
        
        # Close button
        close_btn = ttk.Button(button_frame, text="닫기", command=dialog.destroy)
        close_btn.pack(side='right', padx=5)
        
        # Focus on text widget
        text_widget.focus_set()
        
        # Bind Escape key to close
        dialog.bind('<Escape>', lambda e: dialog.destroy())
        
        # Wait for dialog to close
        dialog.wait_window()

    def on_daily_usage_tree_select(self, event):
        """No-op: details panel removed"""
        pass

    def delete_selected_site(self):
        """Remove the selected site from the suggestion list"""
        site = self.cb_daily_site.get().strip()
        if not site:
            messagebox.showwarning("선택 오류", "삭제할 현장명을 선택해주세요.")
            return
            
        if site in self.sites:
            if messagebox.askyesno("삭제 확인", f"'{site}' 현장명을 목록에서 삭제하시겠습니까?\n(기존 기록은 삭제되지 않습니다.)"):
                self.sites.remove(site)
                self.cb_daily_site['values'] = self.sites
                self.cb_daily_site.set('')
                self.save_tab_config()
                messagebox.showinfo("완료", "현장명이 삭제되었습니다.")
        else:
            messagebox.showinfo("알림", "목록에 없는 현장명입니다.")

    def calculate_rtk_total(self):
        """Calculate total RTK usage"""
        try:
            total = 0
            rtk_categories = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타"]
            for category in rtk_categories:
                value = self.rtk_entries[category].get()
                if value:
                    total += float(value)
            self.rtk_entries["총계"].config(state='normal')
            self.rtk_entries["총계"].delete(0, tk.END)
            # Format as integer if possible, else float
            if total == int(total):
                self.rtk_entries["총계"].insert(0, str(int(total)))
            else:
                self.rtk_entries["총계"].insert(0, f"{total:.1f}")
            self.rtk_entries["총계"].config(state='readonly')
        except ValueError:
            pass  # Ignore invalid input during typing
    
    def format_entry_with_commas(self, event, entry):
        """Autoformat numbers with commas as the user types or leaves the field"""
        try:
            val = entry.get().strip().replace(',', '')
            if not val: return
            
            # Use float or int depending on content
            if '.' in val:
                num = float(val)
                formatted = f"{num:,.1f}"
                if formatted.endswith('.0'): formatted = formatted[:-2]
            else:
                num = int(val)
                formatted = f"{num:,}"
                
            entry.delete(0, tk.END)
            entry.insert(0, formatted)
        except: pass

    def update_daily_test_fee_calc(self, event=None):
        """Auto-calculate Inspection Fee = (Amount * Unit Price) + Travel Expense"""
        try:
            def get_f(entry):
                try:
                    val = entry.get().strip().replace(',', '')
                    return float(val) if val else 0.0
                except: return 0.0

            amount = get_f(self.ent_daily_test_amount)
            price = get_f(self.ent_daily_unit_price)
            travel = get_f(self.ent_daily_travel_cost)
            
            calc_fee = (amount * price) + travel
            
            # Update the fee field with commas
            self.ent_daily_test_fee.delete(0, tk.END)
            self.ent_daily_test_fee.insert(0, f"{calc_fee:,.0f}")
        except:
            pass

    def _add_single_usage_record_logic(self, mat_id, date_val, site, auto_save=True):
        """현장 사용량 기록 1건을 처리하는 핵심 로직 (단일/일괄 공용)"""
        def to_f(ent):
            try:
                # Handle both Entry widgets and raw values
                val = ent.get().replace(',', '') if hasattr(ent, 'get') else str(ent).replace(',', '')
                return float(val) if val else 0.0
            except: return 0.0

        # 1. 작업자 목록 및 시간 데이터 구성
        workers_list = []
        worker_data_map = {}
        for i in range(1, 11):
            group = getattr(self, f'worker_group{i}', None)
            u_key = 'User' if i == 1 else f'User{i}'
            wt_key = 'WorkTime' if i == 1 else f'WorkTime{i}'
            ot_key = 'OT' if i == 1 else f'OT{i}'
            
            if group:
                name = group.get_worker().strip()
                wt = group.get_time().strip()
                ot = group.get_ot().strip()
                worker_data_map[u_key] = name
                worker_data_map[wt_key] = wt
                worker_data_map[ot_key] = ot
                if name: workers_list.append(name)
            else:
                worker_data_map[u_key] = ""
                worker_data_map[wt_key] = ""
                worker_data_map[ot_key] = ""
        
        all_workers = ", ".join(workers_list)

        # 2. RTK 데이터 구성
        rtk_data = {}
        for cat, ent in self.rtk_entries.items():
            rtk_data[f'RTK_{cat}'] = to_f(ent)

        # 3. 차량 데이터 수집 (내용이 있는 상자만 필터링)
        living_boxes = []
        if hasattr(self, 'vehicle_boxes'):
            for b in self.vehicle_boxes:
                if hasattr(b, 'winfo_exists') and b.winfo_exists():
                    v_data = b.get_data()
                    v_no = v_data.get('vehicle_info', '').strip()
                    v_mileage = v_data.get('_raw_mileage', '').strip()
                    # 차량번호, 주행거리, 혹은 점검 항목이 하나라도 기입된 상자만 데이터로 인정
                    has_check = any(v for k, v in v_data.items() if k not in ['vehicle_info', 'mileage', 'remarks', '_raw_mileage'])
                    if v_no or v_mileage or has_check:
                        living_boxes.append(b)

        # 4. 업체(Company) 데이터 수집
        company_data_list = []
        if hasattr(self, 'ndt_company_entries'):
            for company_entries in self.ndt_company_entries:
                company_code = company_entries['_company'].get().strip()
                ndt_data = {name: to_f(company_entries.get(name)) for name in self.ndt_materials_all}
                
                # [REFINEMENT] Create record if there is NDT data OR it is the first entry (guaranteed row)
                if any(v > 0 for v in ndt_data.values()) or company_entries == self.ndt_company_entries[0]:
                    company_data_list.append({
                        '회사코드': company_code,
                        'ndt_data': ndt_data,
                        'is_primary': company_entries == self.ndt_company_entries[0]
                    })
        
        if not company_data_list: # [FALLBACK] Ensure at least one row exists
             company_data_list.append({'회사코드': "", 'ndt_data': {}, 'is_primary': True})

        # 5. 공통 레코드 데이터 구성 (기본값)
        common_data = {
            'Date': date_val,
            'Site': site,
            'MaterialID': mat_id,
            '장비명': self.cb_daily_equip.get().strip(),
            '검사방법': self.cb_daily_test_method.get().strip(),
            '검사품명': self.ent_daily_inspection_item.get().strip(),
            '적용코드': self.ent_daily_applied_code.get().strip(),
            '성적서번호': self.ent_daily_report_no.get().strip(),
            '검사자': "", # Derived from workers list in output
            'Usage': to_f(self.ent_daily_test_amount),
            '검사량': to_f(self.ent_daily_test_amount),
            '단가': to_f(self.ent_daily_unit_price),
            '출장비': to_f(self.ent_daily_travel_cost),
            '업체명': self.cb_daily_company.get().strip(),
            'Unit': self.cb_daily_unit.get().strip(),
        }

        # [NEW] Auto-save new unit, equipment, and site (with improved duplicate check)
        new_unit = common_data['Unit'].strip()
        if new_unit and not any(u.strip() == new_unit for u in self.daily_units):
            self.daily_units.append(new_unit)
            self.refresh_ui_for_list_change('daily_units')

        new_equip = common_data['장비명'].strip()
        if new_equip and not any(e.strip() == new_equip for e in self.equipments):
            self.equipments.append(new_equip)
            self.refresh_ui_for_list_change('equipments')

        new_site = common_data['Site'].strip()
        if new_site and not any(s.strip() == new_site for s in self.sites):
            self.sites.append(new_site)
            self.refresh_ui_for_list_change('sites')

        common_data.update({
            '일식': to_f(self.ent_daily_meal_cost),
            '검사비': to_f(self.ent_daily_test_fee),
            'FilmCount': 0.0,
            'Note': self._get_merged_memo_and_note(),
            'EntryTime': datetime.datetime.now(),
            '회사코드': "",
            **worker_data_map
        })

        # 6. 레코드 생성 (업체 수 vs 차량 수 중 큰 값만큼 생성)
        records_to_save = []
        max_rows = max(len(company_data_list), len(living_boxes))
        
        for i in range(max_rows):
            row_record = common_data.copy()
            
            # --- 업체/수량 데이터 배분 ---
            if i < len(company_data_list):
                cd = company_data_list[i]
                row_record['회사코드'] = cd['회사코드']
                
                # [FIX] Multi-counting Prevention: Only the primary row keeps quantitative data
                if not cd['is_primary']:
                    # Zero out usage/amounts/costs for secondary rows
                    for k in ['Usage', '검사량', '단가', '출장비', '검사비', '일식', 'FilmCount']:
                        row_record[k] = 0.0
                    # Zero out redundant worker quantitative data (WorkTime, OT)
                    # We keep the NAMES (User, User2...) for identification but clear the times
                    for j in range(1, 11):
                        wt_k = 'WorkTime' if j == 1 else f'WorkTime{j}'
                        ot_k = 'OT' if j == 1 else f'OT{j}'
                        row_record[wt_k] = ""
                        row_record[ot_k] = ""
                    # Zero out RTK data
                    for k in rtk_data: row_record[k] = 0.0
                else:
                    # Primary row: keep everything including RTK
                    for k, v in rtk_data.items(): row_record[k] = v

                # Apply NDT data (always row-specific)
                for k, v in cd['ndt_data'].items():
                    row_record[f'NDT_{k}'] = v
            else:
                # Vehicle-only secondary row: Zero out all quantities/costs/times
                for k in ['Usage', '검사량', '단가', '출장비', '검사비', '일식', 'FilmCount']: row_record[k] = 0.0
                for k in rtk_data: row_record[k] = 0.0
                for j in range(1, 11):
                    row_record['WorkTime' if j==1 else f'WorkTime{j}'] = ""
                    row_record['OT' if j==1 else f'OT{j}'] = ""
                for name in self.ndt_materials_all:
                    row_record[f'NDT_{name}'] = 0.0
                row_record['Note'] = "(차량 추가 기록)"

            # --- 차량 데이터 배분 ---
            if i < len(living_boxes):
                v_widget = living_boxes[i]
                v_data = v_widget.get_data()
                v_no = str(v_data.get('vehicle_info', '')).strip()
                v_mileage = str(v_data.get('mileage', '')).strip()
                
                # [SAFEGUARD] Only apply if vehicle number or mileage is present
                if v_no or v_mileage:
                    row_record['차량번호'] = v_no
                    row_record['주행거리'] = v_mileage
                    reserved = ['vehicle_info', 'mileage', 'remarks', '_raw_mileage']
                    # [NEW] Store as category:value|category2:value2 for easier parsing
                    checks_list = []
                    for k, v in v_data.items():
                        if k not in reserved and v:
                            checks_list.append(f"{k}:{v}")
                    row_record['차량점검'] = "|".join(checks_list)
                    row_record['차량비고'] = str(v_data.get('remarks', '')).strip()

                else:
                    row_record['차량번호'] = ""
                    row_record['주행거리'] = ""
                    row_record['차량점검'] = ""
                    row_record['차량비고'] = ""
            else:
                row_record['차량번호'] = ""
                row_record['주행거리'] = ""
                row_record['차량점검'] = ""
                row_record['차량비고'] = ""
            
            records_to_save.append(row_record)

        # 7. 재고 트랜잭션 처리 (원본 로직 유지 - 출고는 실제 물량만큼 한 번만 발생)
        # [NEW] PAUT 및 장비류는 재고 차감에서 제외 (사용자 요청)
        method_raw = common_data.get('검사방법', '')
        method = str(method_raw).strip().upper()
        is_excluded = (method == 'PAUT')
        
        if not is_excluded:
            # 7-1. NDT 약품 차감 (세척제, 현상제 등)
            for cd in company_data_list:
                 if cd.get('ndt_data') and any(v > 0 for v in cd['ndt_data'].values()):
                      self._auto_reconcile_and_register_ndt(date_val, site, cd['ndt_data'], all_workers, cd['회사코드'])
            
            # 7-2. 메인 자재 차감 (RT 필름 등)
            if mat_id and common_data['검사량'] > 0:
                mat_display = self.get_material_display_name(mat_id)
                # [FIX] 소모품인 경우에만 재고 차감 (장비류 제외)
                if self._is_consumable_material(mat_display, method):
                    self._create_manual_stock_transaction(date_val, mat_id, 'OUT', common_data['검사량'], site, all_workers, f"{site} 현장 사용 (자동 차감)")
                else:
                    print(f"[DEBUG] Skipping stock deduction for equipment/excluded: {mat_display}")

        # 8. 데이터프레임 업데이트
        if records_to_save:
            self.daily_usage_df = pd.concat([self.daily_usage_df, pd.DataFrame(records_to_save)], ignore_index=True)
            
            # 리스트 명단 자동 추가 (Master List 관리)
            any_list_changed = False
            if site and site not in self.sites:
                self.sites.append(site); self.sites.sort(); any_list_changed = True
            for w in workers_list:
                if w and w not in self.users:
                    self.users.append(w); self.users.sort(); any_list_changed = True
            
            # 모든 생성된 레코드에서 차량번호 수집 및 추가
            for r in records_to_save:
                v_no = r.get('차량번호', '').strip()
                if v_no and v_no not in self.vehicles:
                    self.vehicles.append(v_no); any_list_changed = True
            
            if any_list_changed:
                self.vehicles.sort()
                self.save_tab_config()

        if auto_save:
            res = self.save_data()
            return len(records_to_save) if res else 0
        return len(records_to_save)

    def add_ndt_company_section(self):
        """회사별 NDT 입력 섹션 추가"""
        if not hasattr(self, 'ndt_company_entries'):
            self.ndt_company_entries = []
        
        company_idx = len(self.ndt_company_entries)
        ndt_materials = self.ndt_materials_all


        
        # Company frame
        company_frame = ttk.LabelFrame(self.ndt_company_container, text=f"회사 #{company_idx + 1}")
        company_frame.pack(fill='x', padx=2, pady=2)
        
        # Company code selector
        header_frame = ttk.Frame(company_frame)
        header_frame.pack(fill='x', padx=2, pady=2)
        ttk.Label(header_frame, text="회사코드:", font=('Arial', 8, 'bold')).pack(side='left', padx=2)
        cb_co = ttk.Combobox(header_frame, width=8, values=getattr(self, 'co_code_list', []))
        cb_co.pack(side='left', padx=2)
        cb_co.set('')  # Default empty
        
        # NDT entries grid
        grid_frame = ttk.Frame(company_frame)
        for c in range(6): grid_frame.columnconfigure(c, weight=1, uniform="ndt_rtk")
        grid_frame.pack(fill='x', padx=2, pady=2)
        
        entries = {'_company': cb_co}  # Store company combobox
        for i, mat in enumerate(ndt_materials):
            r = i // 3
            c = (i % 3) * 2
            ttk.Label(grid_frame, text=f"{mat}:", font=('Arial', 8)).grid(row=r, column=c, padx=1, pady=1, sticky='w')
            e = ttk.Entry(grid_frame, width=6)
            e.grid(row=r, column=c+1, padx=1, pady=1, sticky='ew')
            entries[mat] = e
        
        self.ndt_company_entries.append(entries)
    
    def remove_last_ndt_company(self):
        """마지막 회사 섹션 삭제"""
        if hasattr(self, 'ndt_company_entries') and len(self.ndt_company_entries) > 1:
            # Remove last entry dict
            self.ndt_company_entries.pop()
            # Destroy last widget in container
            widgets = self.ndt_company_container.winfo_children()
            if widgets:
                widgets[-1].destroy()

    def sync_worker_times(self):
        """작업자 1의 설정(주야/작업시간/OT)을 모든 작업자와 동기화"""
        try:
            # Get values from Worker 1
            if not hasattr(self, 'worker_group1'):
                return
                
            master_group = self.worker_group1
            wt1 = master_group.ent_worktime.get().strip()
            ot1 = master_group.ent_ot.get().strip()
            shift1 = master_group.cb_shift.get()
            
            if not wt1 and not ot1:
                messagebox.showwarning("입력 필요", "작업자 1의 작업시간이나 OT를 입력해주세요.")
                return

            for i in range(2, 11):
                group_attr = f'worker_group{i}'
                
                # Check if this worker slot exists
                if not hasattr(self, group_attr):
                    continue
                    
                target_group = getattr(self, group_attr)
                
                # Only apply to workers with names selected
                target_name = target_group.get_worker().strip()
                if not target_name:
                    continue

                # Sync Shift
                target_group.cb_shift.set(shift1)

                # Sync Work Time
                target_group.ent_worktime.set(wt1)
                
                # Sync OT
                target_group.set_ot(ot1)
            
            messagebox.showinfo("완료", "작업자 1의 설정(주야/시간/OT)이 성명이 입력된 모든 작업자에게 적용되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"동기화 중 오류가 발생했습니다: {e}")

    def _load_ndt_product_map(self):
        """config 에서 NDT 약품 → 실제 DB 품목명 매핑 불러오기"""
        try:
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                return cfg.get('ndt_product_map', {})
        except Exception:
            pass
        return {}

    def _save_ndt_product_map(self, map_data):
        """NDT 약품 -> 실제 DB 품목명 매핑 저장"""
        try:
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
            else:
                cfg = {}
            cfg['ndt_product_map'] = map_data
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, indent=4, ensure_ascii=False)
            return True
        except Exception:
            return False

    def open_ndt_product_map_dialog(self):
        """NDT 약품 -> 실제 DB 품목명 매핑 설정 다이얼로그"""
        dlg = tk.Toplevel(self.root)
        dlg.title("NDT 약품-품목 매핑 설정")
        dlg.geometry("500x600")
        dlg.transient(self.root)
        dlg.grab_set()

        main_frame = ttk.Frame(dlg, padding=20)
        main_frame.pack(fill='both', expand=True)

        ttk.Label(main_frame, text="현장 입력 약품명", font=('Arial', 10, 'bold')).grid(row=0, column=0, pady=10)
        ttk.Label(main_frame, text="창고 재고 품목 (매핑)", font=('Arial', 10, 'bold')).grid(row=0, column=1, pady=10)

        ndt_materials = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
        current_map = self._load_ndt_product_map()
        
        material_options = []
        for _, row in self.materials_df.iterrows():
            disp = self.get_material_display_name(row['MaterialID'])
            material_options.append(disp)
        material_options.sort()

        combos = {}
        for i, mat in enumerate(ndt_materials):
            ttk.Label(main_frame, text=mat).grid(row=i+1, column=0, padx=5, pady=5, sticky='w')
            cb = ttk.Combobox(main_frame, values=material_options, width=40)
            cb.grid(row=i+1, column=1, padx=5, pady=5, sticky='ew')
            current_id = current_map.get(mat, "")
            if current_id:
                cb.set(self.get_material_display_name(current_id))
            combos[mat] = cb

        def _save():
            new_map = {}
            for mat, cb in combos.items():
                disp = cb.get().strip()
                if disp:
                    for _, row in self.materials_df.iterrows():
                        if self.get_material_display_name(row['MaterialID']) == disp:
                            new_map[mat] = row['MaterialID']
                            break
            if self._save_ndt_product_map(new_map):
                messagebox.showinfo("성공", "매핑 설정이 저장되었습니다.")
                dlg.destroy()
            else:
                messagebox.showerror("오류", "설정을 저장하지 못했습니다.")

        def _clear():
            for cb in combos.values():
                cb.set('')

        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=len(ndt_materials)+1, column=0, columnspan=2, pady=20)

        ttk.Button(btn_frame, text="저장", command=_save, width=10).pack(side='left', padx=8)
        ttk.Button(btn_frame, text="전체 초기화", command=_clear, width=10).pack(side='left', padx=8)
        ttk.Button(btn_frame, text="닫기", command=dlg.destroy, width=10).pack(side='left', padx=8)

    def open_equipment_management_dialog(self):
        """Dialog to add, edit or delete equipment names in the list"""
        dlg = tk.Toplevel(self.root)
        dlg.title("장비명 목록 관리")
        dlg.geometry("900x700") # Increased size
        dlg.transient(self.root)
        dlg.grab_set()

        main_frame = ttk.Frame(dlg, padding=20)
        main_frame.pack(fill='both', expand=True)

        ttk.Label(main_frame, text="사용자 정의 장비 목록", font=('Malgun Gothic', 12, 'bold')).pack(pady=(0, 5))
        ttk.Label(main_frame, text="(드롭다운 목록에 표시될 기본 장비 명단입니다)", font=('Malgun Gothic', 9)).pack(pady=(0, 10))
        
        # Treeview to show equipment list
        list_container = ttk.Frame(main_frame)
        list_container.pack(fill='both', expand=True)

        # Left side: Current Equipment List
        left_frame = ttk.LabelFrame(list_container, text="현재 장비 목록", padding=5)
        left_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))
        
        columns = ('name',)
        self.equip_tree = ttk.Treeview(left_frame, columns=columns, show='headings', height=15)
        self.equip_tree.heading('name', text='장비명')
        self.equip_tree.column('name', width=350) # Increased width
        
        scrollbar_l_y = ttk.Scrollbar(left_frame, orient="vertical", command=self.equip_tree.yview)
        scrollbar_l_x = ttk.Scrollbar(left_frame, orient="horizontal", command=self.equip_tree.xview)
        self.equip_tree.configure(yscrollcommand=scrollbar_l_y.set, xscrollcommand=scrollbar_l_x.set)
        
        self.equip_tree.pack(side='top', fill='both', expand=True)
        scrollbar_l_y.pack(side='right', fill='y', in_=left_frame)
        scrollbar_l_x.pack(side='bottom', fill='x')

        # Right side: Materials from Stock Status
        right_frame = ttk.LabelFrame(list_container, text="재고현황 품목 (가져오기)", padding=5)
        right_frame.pack(side='right', fill='both', expand=True, padx=(5, 0))

        # Search Bar for Right List
        search_frame = ttk.Frame(right_frame)
        search_frame.pack(fill='x', pady=(0, 5))
        ttk.Label(search_frame, text="🔍 검색:").pack(side='left', padx=2)
        search_var = tk.StringVar()
        search_ent = ttk.Entry(search_frame, textvariable=search_var)
        search_ent.pack(side='left', fill='x', expand=True, padx=2)

        columns_r = ('name', 'sn', 'model')
        self.mat_source_tree = ttk.Treeview(right_frame, columns=columns_r, show='headings', height=15)
        self.mat_source_tree.heading('name', text='품목명')
        self.mat_source_tree.heading('sn', text='SN')
        self.mat_source_tree.heading('model', text='모델명')
        self.mat_source_tree.column('name', width=150)
        self.mat_source_tree.column('sn', width=120)
        self.mat_source_tree.column('model', width=150)

        scrollbar_r_y = ttk.Scrollbar(right_frame, orient="vertical", command=self.mat_source_tree.yview)
        scrollbar_r_x = ttk.Scrollbar(right_frame, orient="horizontal", command=self.mat_source_tree.xview)
        self.mat_source_tree.configure(yscrollcommand=scrollbar_r_y.set, xscrollcommand=scrollbar_r_x.set)
        
        self.mat_source_tree.pack(side='top', fill='both', expand=True)
        scrollbar_r_y.pack(side='right', fill='y', in_=right_frame)
        scrollbar_r_x.pack(side='bottom', fill='x')

        def refresh_tree():
            for item in self.equip_tree.get_children():
                self.equip_tree.delete(item)
            for eq in sorted(self.equipments):
                self.equip_tree.insert('', tk.END, values=(eq,))

        def refresh_mat_tree(*args):
            query = search_var.get().strip().lower()
            for item in self.mat_source_tree.get_children():
                self.mat_source_tree.delete(item)
            
            if not self.materials_df.empty:
                # Group by unique combinations
                unique_items = self.materials_df[['품목명', 'SN', '모델명']].drop_duplicates()
                for _, row in unique_items.iterrows():
                    name = str(row.get('품목명', '')).strip()
                    sn = str(row.get('SN', '')).strip()
                    model = str(row.get('모델명', '')).strip()
                    
                    if name == 'nan': name = ''
                    if sn == 'nan': sn = ''
                    if model == 'nan': model = ''
                    
                    if name or sn or model:
                        # Apply filter
                        if not query or query in name.lower() or query in sn.lower() or query in model.lower():
                            self.mat_source_tree.insert('', tk.END, values=(name, sn, model))

        search_var.trace_add("write", refresh_mat_tree)
        refresh_tree()
        refresh_mat_tree()

        # Action Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=10)

        def add_from_stock():
            selection = self.mat_source_tree.selection()
            if not selection:
                messagebox.showwarning("선택 오류", "가져올 품목을 오른쪽 목록에서 선택해주세요.")
                return
            
            added_count = 0
            for sel in selection:
                vals = self.mat_source_tree.item(sel)['values']
                name = str(vals[0]).strip()
                sn = str(vals[1]).strip()
                model = str(vals[2]).strip()
                
                # Cleanup 'nan' strings
                if name.lower() == 'nan': name = ''
                if sn.lower() == 'nan': sn = ''
                if model.lower() == 'nan': model = ''

                # Create a display string for the equipment list
                # Format: [Model] Name (SN)
                display_parts = []
                
                # Logic: If model exists, add it. If name exists and is different from model, add it.
                if model:
                    display_parts.append(model)
                if name and name != model:
                    display_parts.append(name)
                if sn:
                    display_parts.append(f"({sn})")
                
                display_name = " ".join(display_parts).strip()
                
                if display_name and display_name not in self.equipments:
                    self.equipments.append(display_name)
                    added_count += 1
            
            if added_count > 0:
                refresh_tree()
                self.save_tab_config()
                self.update_material_combo()
                messagebox.showinfo("성공", f"{added_count}개의 품목을 장비 목록에 추가했습니다.")

        def add_equip():
            def save_new():
                name = entry.get().strip()
                if name:
                    if name not in self.equipments:
                        self.equipments.append(name)
                        refresh_tree()
                        refresh_mat_tree()
                        self.save_tab_config()
                        self.update_material_combo()
                        add_win.destroy()
                    else:
                        messagebox.showinfo("중복", "이미 존재하는 장비명입니다.")
                else:
                    messagebox.showwarning("입력 오류", "장비명을 입력해주세요.")

            add_win = tk.Toplevel(dlg)
            add_win.title("장비 추가")
            add_win.geometry("600x150")
            add_win.transient(dlg)
            add_win.grab_set()
            
            ttk.Label(add_win, text="추가할 장비명을 입력하세요:", font=('Malgun Gothic', 10)).pack(pady=10)
            entry = ttk.Entry(add_win, width=70)
            entry.pack(padx=20, pady=5)
            entry.focus_set()
            
            btn_f = ttk.Frame(add_win)
            btn_f.pack(pady=10)
            ttk.Button(btn_f, text="확인", command=save_new, width=10).pack(side='left', padx=5)
            ttk.Button(btn_f, text="취소", command=add_win.destroy, width=10).pack(side='left', padx=5)
            
            add_win.bind('<Return>', lambda e: save_new())

        def delete_equip():
            selection = self.equip_tree.selection()
            if not selection:
                messagebox.showwarning("선택 오류", "삭제할 항목을 선택해주세요.")
                return
            
            count = len(selection)
            if messagebox.askyesno("삭제 확인", f"선택한 {count}개의 장비를 목록에서 삭제하시겠습니까?"):
                for sel in selection:
                    name = self.equip_tree.item(sel)['values'][0]
                    if name in self.equipments:
                        self.equipments.remove(name)
                
                refresh_tree()
                refresh_mat_tree()
                self.save_tab_config()
                self.update_material_combo()

        def edit_equip():
            selection = self.equip_tree.selection()
            if not selection:
                messagebox.showwarning("선택 오류", "수정할 항목을 선택해주세요.")
                return
            
            old_name = self.equip_tree.item(selection[0])['values'][0]
            
            def save_edit():
                new_name = entry.get().strip()
                if new_name and new_name != old_name:
                    if old_name in self.equipments:
                        idx = self.equipments.index(old_name)
                        self.equipments[idx] = new_name
                        refresh_tree()
                        refresh_mat_tree()
                        self.save_tab_config()
                        self.update_material_combo()
                        edit_win.destroy()
                elif new_name == old_name:
                    edit_win.destroy()
                else:
                    messagebox.showwarning("입력 오류", "장비명을 입력해주세요.")

            edit_win = tk.Toplevel(dlg)
            edit_win.title("장비명 수정")
            edit_win.geometry("600x150")
            edit_win.transient(dlg)
            edit_win.grab_set()
            
            ttk.Label(edit_win, text="장비명을 수정하세요:", font=('Malgun Gothic', 10)).pack(pady=10)
            entry = ttk.Entry(edit_win, width=70)
            entry.pack(padx=20, pady=5)
            entry.insert(0, old_name)
            entry.selection_range(0, tk.END)
            entry.focus_set()
            
            btn_f = ttk.Frame(edit_win)
            btn_f.pack(pady=10)
            ttk.Button(btn_f, text="저장", command=save_edit, width=10).pack(side='left', padx=5)
            ttk.Button(btn_f, text="취소", command=edit_win.destroy, width=10).pack(side='left', padx=5)
            
            edit_win.bind('<Return>', lambda e: save_edit())

        ttk.Button(btn_frame, text="◀ 가져오기", command=add_from_stock).pack(side='left', padx=5, expand=True, fill='x')
        ttk.Button(btn_frame, text="➕ 직접추가", command=add_equip).pack(side='left', padx=5, expand=True, fill='x')
        ttk.Button(btn_frame, text="✏️ 수정", command=edit_equip).pack(side='left', padx=5, expand=True, fill='x')
        ttk.Button(btn_frame, text="❌ 삭제", command=delete_equip).pack(side='left', padx=5, expand=True, fill='x')

        ttk.Button(main_frame, text="닫기", command=dlg.destroy).pack(pady=10)

    def open_equipment_search_dialog(self):
        """Dialog to search equipment from master and select it for the form"""
        dlg = tk.Toplevel(self.root)
        dlg.title("장비 검색 및 선택")
        dlg.geometry("700x600")
        dlg.transient(self.root)
        dlg.grab_set()

        main_frame = ttk.Frame(dlg, padding=20)
        main_frame.pack(fill='both', expand=True)

        ttk.Label(main_frame, text="재고현황에서 장비 검색", font=('Malgun Gothic', 12, 'bold')).pack(pady=(0, 10))
        
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill='x', pady=5)
        ttk.Label(search_frame, text="🔍 검색어:").pack(side='left', padx=5)
        search_var = tk.StringVar()
        search_ent = ttk.Entry(search_frame, textvariable=search_var)
        search_ent.pack(side='left', fill='x', expand=True, padx=5)
        search_ent.focus_set()

        # Treeview for results
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill='both', expand=True)
        
        columns = ('name', 'sn', 'model')
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        tree.heading('name', text='품목명')
        tree.heading('sn', text='SN')
        tree.heading('model', text='모델명')
        tree.column('name', width=200)
        tree.column('sn', width=150)
        tree.column('model', width=200)
        
        sb_y = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        sb_x = ttk.Scrollbar(main_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        
        tree.pack(side='left', fill='both', expand=True)
        sb_y.pack(side='right', fill='y')
        sb_x.pack(fill='x')

        def refresh_list(*args):
            query = search_var.get().strip().lower()
            for item in tree.get_children():
                tree.delete(item)
            
            if not self.materials_df.empty:
                # [STRICT] Filter for Active and Non-Consumable (Equipment) items only
                active_mask = pd.to_numeric(self.materials_df.get('Active', 1), errors='coerce').fillna(1) > 0.5
                eq_df = self.materials_df[active_mask]
                
                # Further filter: Exclude consumables
                equipment_df = eq_df[eq_df['품목명'].apply(lambda x: not self._is_consumable_material(str(x).strip(), ''))]
                
                unique_items = equipment_df[['품목명', 'SN', '모델명']].drop_duplicates()
                for _, row in unique_items.iterrows():
                    name = str(row.get('품목명', '')).strip()
                    sn = str(row.get('SN', '')).strip()
                    model = str(row.get('모델명', '')).strip()
                    
                    if name.lower() == 'nan': name = ''
                    if sn.lower() == 'nan': sn = ''
                    if model.lower() == 'nan': model = ''
                    
                    if name or sn or model:
                        if not query or query in name.lower() or query in sn.lower() or query in model.lower():
                            tree.insert('', tk.END, values=(name, sn, model))

        def on_select(e=None):
            selection = tree.selection()
            if not selection: return
            
            vals = tree.item(selection[0])['values']
            name, sn, model = str(vals[0]), str(vals[1]), str(vals[2])
            
            # Format: [Model] Name (SN)
            display_parts = []
            if model and model != 'None' and model != '': display_parts.append(model)
            if name and name != model and name != 'None' and name != '': display_parts.append(name)
            if sn and sn != 'None' and sn != '': display_parts.append(f"({sn})")
            
            display_name = " ".join(display_parts).strip()
            
            if display_name:
                self.cb_daily_equip.delete(0, tk.END)
                self.cb_daily_equip.insert(0, display_name)
                dlg.destroy()

        def on_direct_entry():
            query = search_var.get().strip()
            if not query:
                messagebox.showwarning("입력 오류", "등록할 장비명을 입력해주세요.")
                return
            
            # Use the query as the equipment name directly
            self.cb_daily_equip.delete(0, tk.END)
            self.cb_daily_equip.insert(0, query)
            
            # [SMART] Also add to managed equipment list if not already there
            if query not in self.equipments:
                self.equipments.append(query)
                self.refresh_ui_for_list_change('equipments')
            
            dlg.destroy()

        search_var.trace_add("write", refresh_list)
        tree.bind('<Double-1>', on_select)
        
        refresh_list()

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="선택 완료", command=on_select, width=15).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="입력 명칭으로 신규 등록", command=on_direct_entry, width=22).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="취소", command=dlg.destroy, width=10).pack(side='left', padx=5)

    def open_material_search_dialog(self, target_form='transaction'):
        """Unified dialog to search materials, edit them, or select for forms"""
        dlg = tk.Toplevel(self.root)
        dlg.title("자재 마스터 검색 및 관리")
        dlg.geometry("900x700")
        dlg.transient(self.root)
        dlg.grab_set()

        main_frame = ttk.Frame(dlg, padding=20)
        main_frame.pack(fill='both', expand=True)

        header_text = "전체 재고 자재 검색 및 수정" if target_form == 'registration' else "자재 검색 및 선택"
        ttk.Label(main_frame, text=header_text, font=('Malgun Gothic', 12, 'bold')).pack(pady=(0, 10))
        
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill='x', pady=5)
        ttk.Label(search_frame, text="🔍 검색어:").pack(side='left', padx=5)
        search_var = tk.StringVar()
        search_ent = ttk.Entry(search_frame, textvariable=search_var)
        search_ent.pack(side='left', fill='x', expand=True, padx=5)
        search_ent.focus_set()

        # Treeview for results
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill='both', expand=True)
        
        columns = ('id', 'name', 'model', 'sn', 'spec', 'unit', 'stock')
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        tree.heading('id', text='ID')
        tree.heading('name', text='품목명')
        tree.heading('model', text='모델명')
        tree.heading('sn', text='SN')
        tree.heading('spec', text='규격')
        tree.heading('unit', text='단위')
        tree.heading('stock', text='재고')
        
        tree.column('id', width=50)
        tree.column('name', width=180)
        tree.column('model', width=150)
        tree.column('sn', width=100)
        tree.column('spec', width=100)
        tree.column('unit', width=60)
        tree.column('stock', width=60)
        
        sb_y = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb_y.set)
        
        tree.pack(side='left', fill='both', expand=True)
        sb_y.pack(side='right', fill='y')

        def refresh_list(*args):
            query = search_var.get().strip().lower()
            for item in tree.get_children():
                tree.delete(item)
            
            if not self.materials_df.empty:
                for idx, row in self.materials_df.iterrows():
                    # [STRICT] Filter out inactive items (Active=0) from search dialog
                    # Use a robust check that handles both Series and scalars
                    active_val = row.get('Active', 1)
                    num_active = pd.to_numeric(active_val, errors='coerce')
                    is_active = (num_active > 0.5) if pd.notna(num_active) else True
                    
                    if not is_active:
                        continue
                    
                    name = str(row.get('품목명', '')).strip()
                    sn = str(row.get('SN', '')).strip()
                    model = str(row.get('모델명', '')).strip()
                    spec = str(row.get('규격', '')).strip()
                    unit = str(row.get('관리단위', '')).strip()
                    stock = str(row.get('수량', '0'))
                    # [NEW] Format MaterialID as integer for display
                    mid_raw = row.get('MaterialID', '')
                    mid = ""
                    try:
                        if pd.notna(mid_raw) and str(mid_raw).strip():
                            mid = str(int(float(str(mid_raw).strip())))
                    except:
                        mid = str(mid_raw)
                    
                    if spec == "자동등록": continue
                    
                    # Clean nans
                    name = '' if name.lower() == 'nan' else name
                    sn = '' if sn.lower() == 'nan' else sn
                    model = '' if model.lower() == 'nan' else model
                    spec = '' if spec.lower() == 'nan' else spec
                    
                    if name or sn or model:
                        display_name = name if is_active else f"[비활성] {name}"
                        row_str = f"{name} {sn} {model} {spec} {unit}".lower()
                        if not query or query in row_str:
                            tag = 'active' if is_active else 'inactive'
                            tree.insert('', tk.END, values=(mid, display_name, model, sn, spec, unit, stock), tags=(tag, idx))
                
                tree.tag_configure('inactive', foreground='gray')
                tree.tag_configure('active', foreground='black')

        def on_select(e=None):
            selection = tree.selection()
            if not selection: return
            
            vals = tree.item(selection[0])['values']
            name = str(vals[1])
            model = str(vals[2])
            sn = str(vals[3])
            
            if target_form == 'transaction':
                self.cb_material.set(name)
                self.on_material_selected()
                dlg.destroy()
            elif target_form == 'daily_usage':
                self.cb_daily_material.delete(0, tk.END)
                self.cb_daily_material.insert(0, name)
                dlg.destroy()
            else:
                # Registration form
                self.cb_item_name.set(name)
                if hasattr(self, 'cb_model'): self.cb_model.set(model)
                if hasattr(self, 'ent_sn'): 
                    self.ent_sn.delete(0, tk.END)
                    self.ent_sn.insert(0, sn)
                dlg.destroy()

        def edit_material():
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("선택 오류", "수정할 항목을 선택해주세요.")
                return
            
            item_id = tree.item(selection[0])['values'][0]
            # Find in DF
            mask = self.materials_df['MaterialID'].astype(str) == str(item_id)
            if not mask.any(): return
            
            row_idx = self.materials_df.index[mask][0]
            mat_data = self.materials_df.loc[row_idx]

            edit_win = tk.Toplevel(dlg)
            edit_win.title(f"자재 정보 수정 - ID: {item_id}")
            edit_win.geometry("400x500")
            edit_win.transient(dlg)
            edit_win.grab_set()

            f = ttk.Frame(edit_win, padding=20)
            f.pack(fill='both', expand=True)

            fields = [
                ("품목명", "품목명"), ("모델명", "모델명"), ("SN", "SN"), 
                ("규격", "규격"), ("단위", "관리단위"), ("수량", "수량"), ("재고하한", "재고하한")
            ]
            entries = {}
            for i, (label, col) in enumerate(fields):
                ttk.Label(f, text=f"{label}:").grid(row=i, column=0, padx=5, pady=5, sticky='w')
                ent = ttk.Entry(f, width=30)
                ent.grid(row=i, column=1, padx=5, pady=5, sticky='ew')
                val = str(mat_data.get(col, ''))
                if val.lower() == 'nan': val = ''
                ent.insert(0, val)
                entries[col] = ent

            def save_edit():
                for col, ent in entries.items():
                    val = ent.get().strip()
                    if col in ['수량', '재고하한']:
                        try: val = float(val) if val else 0.0
                        except: val = 0.0
                    self.materials_df.at[row_idx, col] = val
                
                self.save_data()
                self.update_stock_view()
                self.update_material_combo()
                refresh_list()
                messagebox.showinfo("수정 완료", "자재 정보가 업데이트되었습니다.")
                edit_win.destroy()

            ttk.Button(edit_win, text="변경사항 저장", command=save_edit).pack(pady=10)

        def add_new_material():
            """Small dialog to add a missing material quickly"""
            add_win = tk.Toplevel(dlg)
            add_win.title("자재 신규 등록")
            add_win.geometry("450x350")
            add_win.transient(dlg)
            add_win.grab_set()

            form_f = ttk.Frame(add_win, padding=20)
            form_f.pack(fill='both', expand=True)

            entries = {}
            labels = [("품목명", "name"), ("모델명", "model"), ("SN", "sn"), ("규격", "spec"), ("단위", "unit")]
            
            for i, (label, key) in enumerate(labels):
                ttk.Label(form_f, text=f"{label}:").grid(row=i, column=0, padx=5, pady=5, sticky='w')
                ent = ttk.Entry(form_f, width=40)
                ent.grid(row=i, column=1, padx=5, pady=5, sticky='ew')
                entries[key] = ent
                if key == "name":
                    ent.insert(0, search_var.get().strip())
                    ent.focus_set()
                if key == "unit":
                    ent.insert(0, "매")

            def save_new():
                data = {k: e.get().strip() for k, e in entries.items()}
                if not data['name']:
                    messagebox.showwarning("오류", "품목명은 필수입니다.")
                    return
                
                new_id = 10001
                if not self.materials_df.empty:
                    valid_ids = pd.to_numeric(self.materials_df['MaterialID'], errors='coerce').dropna()
                    if not valid_ids.empty:
                        new_id = int(valid_ids.max()) + 1

                new_row = {
                    'MaterialID': new_id,
                    '품목명': data['name'],
                    '모델명': data['model'],
                    'SN': data['sn'],
                    '규격': data['spec'],
                    '관리단위': data['unit'],
                    'Active': 1,
                    '수량': 0,
                    '재고하한': 0
                }
                for col in self.materials_df.columns:
                    if col not in new_row: new_row[col] = ''
                
                self.materials_df = pd.concat([self.materials_df, pd.DataFrame([new_row])], ignore_index=True)
                self.save_data()
                self.update_material_combo()
                self.update_stock_view()
                refresh_list()
                messagebox.showinfo("완료", f"'{data['name']}' 자재가 성공적으로 등록되었습니다.")
                add_win.destroy()

            ttk.Button(add_win, text="등록 완료", command=save_new).pack(pady=10)

        search_var.trace_add("write", refresh_list)
        tree.bind('<Double-1>', on_select)
        refresh_list()

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="✅ 선택 완료", command=on_select, width=15).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="✏️ 정보 수정", command=edit_material, width=15).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="➕ 직접 등록", command=add_new_material, width=15).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="취소", command=dlg.destroy, width=10).pack(side='left', padx=5)

    def open_preferred_item_search_dialog(self):
        """Deprecated in favor of unified open_material_search_dialog"""
        self.open_material_search_dialog(target_form='registration')

    def _get_merged_memo_and_note(self):
        """[NEW] Collects user-entered note and all open memo contents for archiving"""
        main_note = ""
        if hasattr(self, 'ent_daily_note'):
            main_note = self.ent_daily_note.get().strip()
            
        memo_archive = []
        if hasattr(self, 'memos'):
            for key, m in self.memos.items():
                try:
                    title = m['title_entry'].get().strip()
                    content = m['text_widget'].get('1.0', 'end-1c').strip()
                    if content:
                        memo_archive.append(content)
                except: continue
        
        if not memo_archive:
            return main_note
            
        archive_text = " | ".join(memo_archive)
        return f"{main_note} | {archive_text}" if main_note else archive_text

    def load_daily_usage_to_form(self, record):
        """[NEW] Load a database record back into the input form for editing or report generation"""
        try:
            def to_f_local(v):
                try: return float(str(v).replace(',', '')) if v else 0.0
                except: return 0.0

            # 1. Basic Info
            if 'Date' in record:
                try: self.ent_daily_date.set_date(pd.to_datetime(record['Date']))
                except: pass
            if '업체명' in record:
                self.cb_daily_company.set(str(record['업체명']).split(' [')[0])
            if 'Site' in record:
                self.cb_daily_site.set(str(record['Site']))
            if '적용코드' in record:
                self.ent_daily_applied_code.delete(0, tk.END)
                self.ent_daily_applied_code.insert(0, str(record['적용코드']))
            if '장비명' in record:
                self.cb_daily_equip.delete(0, tk.END)
                self.cb_daily_equip.insert(0, str(record['장비명']))
            if '성적서번호' in record:
                self.ent_daily_report_no.delete(0, tk.END)
                self.ent_daily_report_no.insert(0, str(record['성적서번호']))
            if '검사품명' in record:
                self.ent_daily_inspection_item.delete(0, tk.END)
                self.ent_daily_inspection_item.insert(0, str(record['검사품명']))
            if '검사방법' in record:
                self.cb_daily_test_method.set(str(record['검사방법']))
            if 'Unit' in record:
                self.cb_daily_unit.set(str(record['Unit']))
            elif '단위' in record:
                self.cb_daily_unit.set(str(record['단위']))
            
            # 2. Quantities & Costs
            def set_val(ent, key):
                val = record.get(key, 0)
                ent.delete(0, tk.END)
                if isinstance(val, (int, float)):
                    if val != 0: ent.insert(0, f"{val:,.0f}" if any(x in key for x in ['비', '가', '일식']) else str(val))
                else: 
                    ent.insert(0, str(val))

            set_val(self.ent_daily_test_amount, '검사량')
            set_val(self.ent_daily_unit_price, '단가')
            set_val(self.ent_daily_travel_cost, '출장비')
            set_val(self.ent_daily_test_fee, '검사비')
            set_val(self.ent_daily_meal_cost, '일식')
            
            # 3. Material
            if 'MaterialID' in record:
                disp_name = self.get_material_display_name(record['MaterialID'])
                if isinstance(self.cb_daily_material, ttk.Combobox):
                    self.cb_daily_material.set(disp_name)
                else:
                    self.cb_daily_material.delete(0, tk.END)
                    self.cb_daily_material.insert(0, disp_name)

            # 4. Workers
            # First, clear existing workers if needed or just overwrite
            for i in range(1, 11):
                group = getattr(self, f'worker_group{i}', None)
                if not group: continue
                u_key = 'User' if i == 1 else f'User{i}'
                wt_key = 'WorkTime' if i == 1 else f'WorkTime{i}'
                ot_key = 'OT' if i == 1 else f'OT{i}'
                
                group.cb_name.set(self.clean_nan(record.get(u_key, '')))
                # [FIX] Use set_time to properly parse (Shift) Time string and populate both shift and time widgets
                group.set_time(self.clean_nan(record.get(wt_key, '')))
                
                group.ent_ot.delete(0, tk.END)
                group.ent_ot.insert(0, self.clean_nan(record.get(ot_key, '')))

            # 5. RTK

            for cat, ent in self.rtk_entries.items():
                if cat == "총계": continue
                val = record.get(f'RTK_{cat}', 0)
                ent.delete(0, tk.END)
                if abs(to_f_local(val)) > 0.001: ent.insert(0, str(val))
            self.calculate_rtk_total()


            # 6. Vehicle
            v_no = self.clean_nan(record.get('차량번호', ''))
            if v_no:
                if not hasattr(self, 'vehicle_boxes') or not self.vehicle_boxes:
                    self.add_vehicle_inspection_box()
                
                v_widget = self.vehicle_boxes[0]
                v_insp_raw = str(record.get('차량점검', ''))
                v_parsed = {'vehicle_info': v_no, 'mileage': self.clean_nan(record.get('주행거리', '')), 'remarks': self.clean_nan(record.get('차량비고', ''))}
                if ':' in v_insp_raw:
                    # New format
                    for pair in v_insp_raw.split('|'):
                        if ':' in pair:
                            k, v = pair.split(':', 1)
                            v_parsed[k] = v
                elif ',' in v_insp_raw or v_insp_raw:
                    # Old format (keys only)
                    for k in v_insp_raw.split(','):
                        k_clean = k.strip()
                        if k_clean:
                            # Map legacy True to a default "positive" value (since only checked items were saved)
                            if 'locking' in k_clean: v_parsed[k_clean] = '잠금'
                            elif 'cleaning' in k_clean: v_parsed[k_clean] = '함'
                            else: v_parsed[k_clean] = '양호'
                v_widget.set_data(v_parsed)
            
            # 7. NDT Chemicals
            if hasattr(self, 'ndt_company_entries') and self.ndt_company_entries:
                # [STABILITY] Clear first if multiple companies, but for now focus on the primary/first entry
                first = self.ndt_company_entries[0]
                ndt_names = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
                for name in ndt_names:
                    col_name = f"NDT_{name}"
                    if col_name in record:
                        val = record[col_name]
                        if name in first:
                            first[name].delete(0, tk.END)
                            if abs(to_f_local(val)) > 0.001:
                                first[name].insert(0, str(val))
                
                # Also load company if available
                if '회사코드' in record:
                    first['_company'].set(str(record['회사코드']))

            # 8. Note
            if 'Note' in record:
                self.ent_daily_note.delete(0, tk.END)
                self.ent_daily_note.insert(0, str(record['Note']))

        except Exception as e:
            print(f"Error loading record to form: {e}")

    def clear_daily_usage_form_all(self, keep_date=False):
        """현장별 일일 사용량 입력 폼의 모든 필드를 초기화합니다."""
        # 1. 콤보박스 선택 해제
        for cb in [self.cb_daily_company, self.cb_daily_site, 
                   self.cb_daily_test_method, self.cb_daily_unit]:
            cb.set('')
        
        # 2. 날짜 초기화 (연속 기입 시 편의를 위해 keep_date 옵션 제공)
        if not keep_date:
            try:
                self.ent_daily_date.set_date(datetime.date.today())
            except: pass
        
        # 3. 모든 Entry 필드 비우기
        for ent in [self.ent_daily_inspection_item, self.ent_daily_applied_code, 
                   self.ent_daily_equip, self.cb_daily_material, # Both are now Entries
                   self.ent_daily_test_amount, self.ent_daily_unit_price, 
                   self.ent_daily_report_no, self.ent_daily_travel_cost, 
                   self.ent_daily_meal_cost, self.ent_daily_test_fee, self.ent_daily_note]:
            ent.delete(0, tk.END)
            
        # 기본값 복구
        self.cb_daily_unit.set('매')
        
        # 4. NDT 자재 및 RTK 필드 초기화
        if hasattr(self, 'ndt_company_entries'):
            for entries in self.ndt_company_entries:
                for mat in ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]:
                    if mat in entries: entries[mat].delete(0, tk.END)
        
        for ent in self.rtk_entries.values():
            if ent.winfo_exists():
                ent.config(state='normal')
                ent.delete(0, tk.END)
                if ent == self.rtk_entries.get("총계"): 
                    ent.config(state='readonly')
        
        # 5. 차량 점검 섹션 초기화
        if hasattr(self, 'vehicle_boxes'):
            for box in self.vehicle_boxes:
                if box.winfo_exists() and hasattr(box, 'reset_fields'):
                    box.reset_fields()
                    
        # 6. 작업자 섹션 초기화
        for i in range(1, 11):
            group = getattr(self, f'worker_group{i}', None)
            if group:
                group.cb_name.set('')
                group.cb_shift.set('주간')
                group.ent_worktime.set('')
                group.ent_ot.delete(0, tk.END)

    def add_daily_usage_entry(self):
        """현장별 일일 사용량 단일 등록 (리팩토링 버전)"""
        try:
            # 1. 기본 정보 및 유효성 검사
            date_val = self.ent_daily_date.get_date()
            site = self.cb_daily_site.get().strip()
            if not site:
                messagebox.showwarning("입력 오류", "현장명을 입력해주세요.")
                return

            # 2. 품목 ID 확인 (디스플레이 명칭 기반 매칭)
            mat_display = self.cb_daily_material.get().strip()
            mat_id = ""
            if mat_display:
                for _, row in self.materials_df.iterrows():
                    if self.get_material_display_name(row['MaterialID']) == mat_display:
                        mat_id = row['MaterialID']
                        # 휴면 계정 자동 활성화
                        if row.get('Active', 1) == 0:
                            self.materials_df.loc[self.materials_df['MaterialID'] == mat_id, 'Active'] = 1
                        break
                
                # [FIX] Only auto-register if it's a CONSUMABLE (Drug/Film). Equipment should NOT be registered.
                if not mat_id and mat_display:
                    if self._is_consumable_material(mat_display, self.cb_daily_test_method.get().strip()):
                        mat_id = self.register_new_material(mat_display, warehouse='현장', 규격='자동등록')
                    else:
                        # For equipment or non-consumables, just store the name as ID (skips stock tracking)
                        mat_id = mat_display

            # 3. 중복 저장 방지 체크
            if not self.daily_usage_df.empty:
                try:
                    # Normalize date for comparison
                    check_date = pd.to_datetime(date_val).date()
                    # Ensure MaterialID is matched correctly (exact string match)
                    existing = self.daily_usage_df[
                        (pd.to_datetime(self.daily_usage_df['Date']).dt.date == check_date) & 
                        (self.daily_usage_df['Site'] == site) & 
                        (self.daily_usage_df['MaterialID'] == mat_id) &
                        (self.daily_usage_df['검사품명'].astype(str).str.strip() == self.ent_daily_inspection_item.get().strip()) &
                        (self.daily_usage_df['적용코드'].astype(str).str.strip() == self.ent_daily_applied_code.get().strip())
                    ]
                    if not existing.empty:
                        if not messagebox.askyesno("중복 확인", 
                            f"이미 {date_val} 날짜에 '{site}' 현장의 '{mat_display}' 품목 기록이 {len(existing)}건 존재합니다.\n"
                            "동일한 설정으로 추가 저장하시겠습니까?"):
                            return
                except Exception as e:
                    print(f"DEBUG: Duplicate check failed: {e}")

            # 4. 핵심 로직 실행 (단건 저장)
            saved_count = self._add_single_usage_record_logic(mat_id, date_val, site, auto_save=True)
            if saved_count > 0:
                # [SMART VISIBILITY] Adjust filters to ensure the NEW record is visible
                try:
                    current_start = self.ent_daily_start_date.get_date()
                    current_end = self.ent_daily_end_date.get_date()
                    
                    if date_val < current_start:
                        self.ent_daily_start_date.set_date(date_val)
                    if date_val > current_end:
                        self.ent_daily_end_date.set_date(date_val)
                except: pass
                
                # 성공 시 필드 초기화 및 뷰 갱신
                if hasattr(self, 'ndt_company_entries'):
                    for company_entries in self.ndt_company_entries:
                        for name in self.ndt_materials_all:
                            if name in company_entries:
                                company_entries[name].delete(0, tk.END)
                
                # RTK 필드 초기화
                for cat, ent in self.rtk_entries.items():
                    if cat != "총계":
                        ent.delete(0, tk.END)
                self.rtk_entries["총계"].config(state='normal')
                self.rtk_entries["총계"].delete(0, tk.END)
                self.rtk_entries["총계"].config(state='readonly')
                
                # [V13_RESET_REQUESTED] Clearing fields after successful entry based on user request
                # [FIX] Keep the current date for consecutive entries on the same day
                self.clear_daily_usage_form_all(keep_date=True)

                # Reset focus to first logical field (Company)
                self.cb_daily_company.focus_set()

                # [V13.1] Keep workers too for consecutive entries
                # for i in range(1, 11):
                #     group = getattr(self, f'worker_group{i}', None)
                #     if group:
                #         group.cb_name.set('')
                #         group.cb_shift.set('')
                #         group.ent_worktime.set('')
                #         group.ent_ot.delete(0, tk.END)

                # NDT 섹션 초기화
                if hasattr(self, 'ndt_company_entries'):
                    while len(self.ndt_company_entries) > 1:
                        self.remove_last_ndt_company()
                    if self.ndt_company_entries:
                        first = self.ndt_company_entries[0]
                        first['_company'].set('')
                        for k in self.ndt_materials_all:
                            if k in first: first[k].delete(0, tk.END)

                # 차량 점검 필드 초기화 (safeguarded)
                # [NEW] 현장별 저장 시 차량, 목록, 메모 레이블창 자동 닫기 (User request)
                for key in list(getattr(self, 'vehicle_inspections', {}).keys()):
                    self.remove_box(key)
                for key in list(getattr(self, 'checklists', {}).keys()):
                    self.remove_box(key)
                for key in list(getattr(self, 'memos', {}).keys()):
                    self.remove_box(key)

                self.update_daily_usage_view()
                self.refresh_inquiry_filters()
                self.update_stock_view()
                
                # [NEW] Scroll to the bottom to see new entries
                if hasattr(self, 'daily_usage_tree'):
                    self.daily_usage_tree.yview_moveto(1.0)
                
                messagebox.showinfo("성공", f"{saved_count}건의 기록이 안전하게 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"기록 저장 중 오류 발생: {e}")

    def load_report_mapping(self):
        """Load report mapping from JSON file or return defaults"""
        mapping_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'resources', 'report_mapping.json')
        if not os.path.exists(mapping_path):
            # Return defaults matching the manager's default_mapping
            return {
                'header': {'date': 'F2'},
                'general': {
                    'company': 'E5', 'project_name': 'E6', 'standard': 'E7', 'equipment': 'E8',
                    'report_no': 'K5', 'inspection_item': 'K6', 'inspector': 'K7', 'car_no': 'K8'
                },
                'methods': {
                    'RT': {'row': '12'}, 'UT': {'row': '13'}, 'MT': {'row': '14'}, 'PT': {'row': '15'},
                    'HT': {'row': '16'}, 'VT': {'row': '17'}, 'LT': {'row': '18'}, 'ET': {'row': '19'},
                    'PAUT': {'row': '20'}
                },
                'rtk': {
                    'center_miss': 'C32', 'density': 'E32', 'marking_miss': 'G32', 'film_mark': 'I32',
                    'handling': 'K32', 'customer_complaint': 'M32', 'etc': 'O32', 'total': 'Q32'
                },
                'ot': {
                    'row1_name': 'B38', 'row1_hours': 'K38',
                    'row2_name': 'B39', 'row2_hours': 'K39'
                },
                'materials': {
                    'RT T200': '43', 'RT AA400': '44', 'MT WHITE': '46', 'MT 7C-BLACK': '47',
                    'PT Penetrant': '48', 'PT Cleaner': '49', 'PT Developer': '50'
                }
            }
        try:
            with open(mapping_path, 'r', encoding='utf-8') as f:
                import json
                mapping = json.load(f)
                
                # Robust cleaning of method rows (e.g. 'E13' -> 13)
                if 'methods' in mapping:
                    import re
                    for m_key, m_val in mapping['methods'].items():
                        if isinstance(m_val, dict) and 'row' in m_val:
                            row_str = str(m_val['row'])
                            match = re.search(r'\d+', row_str)
                            if match: m_val['row'] = int(match.group())
                return mapping
        except:
            return {}

    def save_report_mapping(self, mapping):
        """Save report mapping to JSON file"""
        resources_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'resources')
        if not os.path.exists(resources_dir): 
            try: os.makedirs(resources_dir)
            except: pass
        mapping_path = os.path.join(resources_dir, 'report_mapping.json')
        try:
            print(f"DEBUG: Saving report mapping to {mapping_path}")
            with open(mapping_path, 'w', encoding='utf-8') as f:
                json.dump(mapping, f, indent=4, ensure_ascii=False)
            return True
        except Exception as e:
            print(f"DEBUG: Save failed: {e}")
            messagebox.showerror("저장 오류", f"매핑 설정을 저장하지 못했습니다: {e}")
            return False

    def open_report_mapping_dialog(self):
        """Open a dialog to configure Excel mapping for Daily Work Report"""
        dialog = tk.Toplevel(self.root)
        dialog.title("작업일보 엑셀 매핑 설정")
        dialog.geometry("600x800")
        dialog.transient(self.root)
        dialog.grab_set()

        mapping = self.load_report_mapping()
        
        main_frame = ttk.Frame(dialog, padding=10)
        main_frame.pack(fill='both', expand=True)

        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill='both', expand=True)

        # Helper to create grid of entries
        def create_entry_grid(parent, fields, current_data):
            entries = {}
            for i, (label_text, key) in enumerate(fields):
                ttk.Label(parent, text=label_text).grid(row=i, column=0, padx=5, pady=5, sticky='w')
                ent = ttk.Entry(parent, width=15)
                # Ensure we handle missing keys or nested dicts in current_data
                val = current_data.get(key, '')
                if isinstance(val, dict): val = val.get('row', '')
                ent.insert(0, str(val))
                ent.grid(row=i, column=1, padx=5, pady=5, sticky='w')
                entries[key] = ent
            return entries

        # Tab 1: General & Header
        tab1 = ttk.Frame(notebook, padding=10)
        notebook.add(tab1, text="일반/상단")
        gen_fields = [
            ("날짜 (Date)", "date_header"),
            ("업체명 (Company)", "company"), ("공사명 (Project)", "project_name"),
            ("적용규격 (Standard)", "standard"), ("장비명 (Equipment)", "equipment"),
            ("성적서번호 (Report No)", "report_no"), ("검사품명 (Item)", "inspection_item"),
            ("검사자 (Inspector)", "inspector"), ("차량번호 (Car No)", "car_no")
        ]
        tab1_data = mapping.get('general', {}).copy()
        tab1_data['date_header'] = mapping.get('header', {}).get('date', 'F2')
        tab1_entries = create_entry_grid(tab1, gen_fields, tab1_data)

        # Tab 2: Methods (Row Numbers)
        tab2 = ttk.Frame(notebook, padding=10)
        notebook.add(tab2, text="검사공법(행)")
        method_keys = ["RT", "UT", "MT", "PT", "HT", "VT", "LT", "ET", "PAUT"]
        method_fields = [(f"{m} (시작 행 번호)", m) for m in method_keys]
        tab2_entries = create_entry_grid(tab2, method_fields, mapping.get('methods', {}))

        # Tab 3: RTK Results
        tab3 = ttk.Frame(notebook, padding=10)
        notebook.add(tab3, text="불량율(RTK)")
        rtk_keys = ['center_miss', 'density', 'marking_miss', 'film_mark', 'handling', 'customer_complaint', 'etc', 'total']
        rtk_fields = [
            ("센터미스", "center_miss"), ("농도", "density"), ("마킹미스", "marking_miss"),
            ("필름마크", "film_mark"), ("취급부주의", "handling"), ("고객불만", "customer_complaint"),
            ("기타", "etc"), ("총계", "total")
        ]
        tab3_entries = create_entry_grid(tab3, rtk_fields, mapping.get('rtk', {}))

        # Tab 4: Materials (Row Numbers)
        tab4 = ttk.Frame(notebook, padding=10)
        notebook.add(tab4, text="자재(행)")
        mat_keys = list(mapping.get('materials', {}).keys())
        if not mat_keys: mat_keys = ['RT T200', 'RT AA400', 'MT WHITE', 'MT 7C-BLACK', 'PT Penetrant', 'PT Cleaner', 'PT Developer']
        mat_fields = [(f"{m} (행 번호)", m) for m in mat_keys]
        tab4_entries = create_entry_grid(tab4, mat_fields, mapping.get('materials', {}))

        def save_and_close():
            try:
                new_mapping = {
                    'header': {'date': tab1_entries['date_header'].get().strip()},
                    'general': {k: tab1_entries[k].get().strip() for k in ['company', 'project_name', 'standard', 'equipment', 'report_no', 'inspection_item', 'inspector', 'car_no']},
                    'methods': {m: {'row': tab2_entries[m].get().strip()} for m in method_keys},
                    'rtk': {k: tab3_entries[k].get().strip() for k in rtk_keys},
                    'ot': mapping.get('ot', {
                        'row1_name': 'B38', 'row1_hours': 'K38',
                        'row2_name': 'B39', 'row2_hours': 'K39'
                    }),
                    'materials': {k: tab4_entries[k].get().strip() for k in mat_keys}
                }
                
                if self.save_report_mapping(new_mapping):
                    messagebox.showinfo("저장 완료", "매핑 설정이 저장되었습니다.")
                    dialog.destroy()
            except Exception as e:
                messagebox.showerror("오류", f"설정 구성 중 오류가 발생했습니다: {e}")

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=10)
        ttk.Button(btn_frame, text="✅ 설정 저장", command=save_and_close).pack(side='right', padx=5)
        ttk.Button(btn_frame, text="❌ 취소", command=dialog.destroy).pack(side='right', padx=5)

    def export_daily_work_report(self):
        """작업일보를 엑셀 템플릿에 출력합니다."""
        try:
            template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'resources', 'Template_DailyWorkReport.xlsx')
            if not os.path.exists(template_path):
                template_path = r'c:\Users\jjch2\Desktop\보고서Project PROVIDENCE\Request\PMI\Na-aba\home\resources\Template_DailyWorkReport.xlsx'
            
            if not os.path.exists(template_path):
                messagebox.showerror("오류", "작업일보 템플릿(Template_DailyWorkReport.xlsx)을 찾을 수 없습니다.")
                return

            def _clean_str(v, default=''):
                if v is None: return default
                s = str(v).strip()
                if s.lower() in ('nan', 'none', ''): return default
                return s

            date_val = self.ent_daily_date.get_date()
            site = self.cb_daily_site.get().strip()
            method = self.cb_daily_test_method.get().strip() # Default method from UI
            
            data = {
                'date': date_val,
                'company': _clean_str(self.cb_daily_company.get(), '원자력건설'),
                'project_name': site,
                'standard': _clean_str(self.ent_daily_applied_code.get(), 'KS'),
                'equipment': _clean_str(self.cb_daily_equip.get()),
                'report_no': _clean_str(self.ent_daily_report_no.get()), 
                'inspection_item': _clean_str(self.ent_daily_inspection_item.get()), 
                'inspector': '', 
                'car_no': '', 
                'methods': {},
                'rtk': {},
                'ot_status': [],
                'materials': {}
            }

            if hasattr(self, 'ndt_company_entries') and self.ndt_company_entries:
                company = self.ndt_company_entries[0].get('_company', tk.Variable()).get().strip()
                if company: data['company'] = company

            # DB 데이터 집합 (Option 1: 전체 합산 버전)
            site_records = pd.DataFrame()
            if not self.daily_usage_df.empty:
                try:
                    df_copy = self.daily_usage_df.copy()
                    
                    # [SMART SELECTION] 리스트에서 여러 개를 선택했는지 확인
                    selection = self.daily_usage_tree.selection()
                    if len(selection) > 1:
                        # 여러 개 선택된 경우, 선택된 항목들만 모아서 출력
                        selected_indices = []
                        for item in selection:
                            tags = self.daily_usage_tree.item(item, 'tags')
                            if tags and tags[0].isdigit():
                                idx = int(tags[0])
                                if idx in df_copy.index:
                                    selected_indices.append(idx)
                        if selected_indices:
                            site_records = df_copy.loc[selected_indices]
                            # 현장명/날짜는 첫 번째 선택된 항목 기준으로 업데이트 (데이터 정합성 보장)
                            if 'Site' in site_records.columns: site = str(site_records.iloc[0]['Site']).strip()
                            if 'Date' in site_records.columns: date_val = pd.to_datetime(site_records.iloc[0]['Date']).date()
                        
                    if site_records.empty:
                        # 하나만 선택되었거나 선택이 없는 경우: 현재 폼의 날짜/현장 기준으로 전체 합산
                        # [ROBUST MATCHING] 날짜와 현장명 비교 시 공백 등 무시하여 정확도 향상
                        check_date = pd.to_datetime(date_val).date()
                        site_col = 'Site' if 'Site' in df_copy.columns else '현장' if '현장' in df_copy.columns else ''
                        date_col = 'Date' if 'Date' in df_copy.columns else '날짜' if '날짜' in df_copy.columns else ''
                        
                        if site_col and date_col:
                            df_copy['Date_norm'] = pd.to_datetime(df_copy[date_col], errors='coerce').dt.date
                            site_records = df_copy[
                                (df_copy['Date_norm'] == check_date) & 
                                (df_copy[site_col].astype(str).str.strip().str.upper() == str(site).strip().upper())
                            ]
                    
                    # [NEW] site_records에서 장비명 및 업체명 집계 업데이트
                    if not site_records.empty:
                        if '장비명' in site_records.columns:
                            equips = [_clean_str(x) for x in site_records['장비명'].dropna().unique() if _clean_str(x)]
                            if equips: data['equipment'] = ", ".join(equips)
                        
                        comp_col = '업체명' if '업체명' in site_records.columns else ''
                        if comp_col:
                            comps = [_clean_str(x) for x in site_records[comp_col].dropna().unique() if _clean_str(x)]
                            if comps: data['company'] = comps[0] # 첫 번째 유효한 업체명 사용
                except Exception as e:
                    print(f"[DEBUG] Site Records Collection Error: {e}")
                    pass

            all_vehicles = []
            if hasattr(self, 'vehicle_boxes'):
                for box in self.vehicle_boxes:
                    v = box.cb_vehicle_info.get().strip()
                    if v and v not in all_vehicles: all_vehicles.append(v)
            if not site_records.empty and '차량번호' in site_records.columns:
                for v in site_records['차량번호'].dropna().unique():
                    v_str = str(v).strip()
                    if v_str and v_str not in all_vehicles: all_vehicles.append(v_str)
            data['car_no'] = ", ".join(all_vehicles)

            # [NEW] 공사 수행현황 (Section 1) - DB site_records에 있는 모든 방식 집계
            method_col = '검사방법' if '검사방법' in site_records.columns else 'TestMethod' if 'TestMethod' in site_records.columns else ''
            
            if not site_records.empty and method_col:
                for m_name, m_group in site_records.groupby(method_col):
                    if not m_name or str(m_name).lower() == 'nan': continue
                    m_name_str = str(m_name)
                    
                    qty_sum = pd.to_numeric(m_group['Usage'], errors='coerce').fillna(0).sum()
                    price_max = pd.to_numeric(m_group['단가'], errors='coerce').fillna(0).max()
                    travel_sum = pd.to_numeric(m_group['출장비'], errors='coerce').fillna(0).sum()
                    total_sum = pd.to_numeric(m_group['검사비'], errors='coerce').fillna(0).sum()
                    
                    # 단위 정보 찾기 (Unit 또는 단위 컬럼)
                    unit_col = 'Unit' if 'Unit' in m_group.columns else '단위' if '단위' in m_group.columns else ''
                    unit_val = str(m_group.iloc[0].get(unit_col, '매')) if unit_col else '매'
                    
                    data['methods'][m_name_str] = {
                        'unit': unit_val,
                        'qty': qty_sum,
                        'price': price_max,
                        'travel': travel_sum,
                        'total': total_sum
                    }
            else:
                # Fallback to UI values if DB is empty or column missing
                method = self.cb_daily_test_method.get().strip()
                unit_val = self.cb_daily_unit.get().strip() 
                qty_val = self.ent_daily_test_amount.get().strip()
                price_val = self.ent_daily_unit_price.get().strip()
                travel_val = self.ent_daily_travel_cost.get().strip()
                total_val = self.ent_daily_test_fee.get().strip()
                
                if method:
                    data['methods'][method] = {
                        'unit': unit_val,
                        'qty': float(qty_val.replace(',', '')) if qty_val else 0,
                        'price': float(price_val.replace(',', '')) if price_val else 0,
                        'travel': float(travel_val.replace(',', '')) if travel_val else 0,
                        'total': float(total_val.replace(',', '')) if total_val else 0
                    }

            # 작업자 및 O/T (간소화 버전)
            # [NEW] 작업자 / OT 정보 DB에서 집계 (현장 탭 기록 기준)
            def _clean_name(n):
                if not n or str(n).lower() in ('nan', 'none', ''): return ''
                text = str(n).strip()
                titles = ['부장', '차장', '과장', '대리', '주임', '기사', '선임', '수석', '책임',
                          '팀장', '이사', '본부장', '실장', '소장', '직장', '반장', '팀원', '계장']
                for t in titles:
                    import re as _re
                    text = _re.sub(r'[\s/(\[]*' + t + r'[\s)\]]*$', '', text)
                    text = _re.sub(r'^[\s/(\[]*' + t + r'[\s)\]]*', '', text)
                return text.strip()

            inspectors = []
            ot_groups = {}  # key: (work_time, ot_amount) -> {names:[], company:''}

            if not site_records.empty:
                company_val = data.get('company', '')
                for _, row in site_records.iterrows():
                    for i in range(1, 11):
                        u_key   = 'User'     if i == 1 else f'User{i}'
                        wt_key  = 'WorkTime' if i == 1 else f'WorkTime{i}'
                        ot_key  = 'OT'       if i == 1 else f'OT{i}'

                        name = str(row.get(u_key, '')).strip()
                        if not name or name == 'nan': continue

                        if name not in inspectors:
                            inspectors.append(name)

                        wt  = str(row.get(wt_key, '')).strip()
                        ot_raw = str(row.get(ot_key,  '')).strip()
                        
                        # [FIX] Handle decimals correctly to prevent 26000.0 becoming 260000
                        try:
                            # Try parsing as float first to handle .0
                            oa_num = float(ot_raw.replace(',', '')) if ot_raw and ot_raw != 'nan' else 0
                            oa = str(int(oa_num)) if oa_num > 0 else ''
                        except:
                            # Fallback to digit-only if float parsing fails
                            oa = ''.join(c for c in ot_raw if c.isdigit())

                        if wt == 'nan': wt = ''
                        if not wt and not oa: continue

                        key = (wt, oa)
                        if key not in ot_groups:
                            ot_groups[key] = {'names': [], 'company': company_val}
                        if name not in ot_groups[key]['names']:
                            ot_groups[key]['names'].append(name)

            # Inspector display (titles always stripped)
            disp_insp = [_clean_name(n) for n in inspectors]
            if len(disp_insp) > 3:
                data['inspector'] = ', '.join(disp_insp[:3]) + f" 외 {len(disp_insp) - 3}명"
            else:
                data['inspector'] = ', '.join(disp_insp)

            # Build ot_status list
            data['ot_status'] = []
            if ot_groups:
                for (wt, oa), grp in ot_groups.items():
                    names = grp['names']
                    name_disp = ', '.join([_clean_name(n) for n in names])
                    wt_disp = wt
                    
                    # Match method for this group if possible
                    curr_method = method
                    if not site_records.empty and method_col:
                        for _, r_match in site_records.iterrows():
                            if any(str(r_match.get(f'User{k}' if k>1 else 'User', '')).strip() in names for k in range(1, 11)):
                                m_found = str(r_match.get(method_col, '')).strip()
                                if m_found and m_found.lower() != 'nan':
                                    curr_method = m_found; break

                    data['ot_status'].append({
                        'names':      name_disp,
                        'ot_hours':   wt_disp,
                        'ot_amount':  f'{int(oa):,}' if oa else '',
                        'company':    grp['company'],
                        'method':     curr_method,
                    })
            else:
                # Fallback to UI worker groups if DB is empty
                for i in range(1, 11):
                    group = getattr(self, f'worker_group{i}', None)
                    if group:
                        w_name = group.cb_name.get().strip()
                        w_wt = group.ent_worktime.get().strip()
                        w_ot = group.ent_ot.get().strip()
                        if w_name and (w_wt or w_ot):
                            data['ot_status'].append({
                                'names':      _clean_name(w_name),
                                'ot_hours':   w_wt,
                                'ot_amount':  w_ot,
                                'company':    data.get('company', ''),
                                'method':     method,
                            })

            

            # 2.5 차량 및 안전 점검 수집 (섹션 3)

            data['vehicles'] = []
            if hasattr(self, 'vehicle_boxes'):
                for box in self.vehicle_boxes:
                    data['vehicles'].append(box.get_data())

            # 자재 정보 수집 (NDT 섹션)
            data['selected_material'] = self.cb_daily_material.get().strip()

            # [NEW] RT 품목별 수량 - DB site_records의 MaterialID로 그룹화 + 품목명/규격 조회
            # D열: 품목명(Name), F열: 규격(Spec), M열: 사용수량(Usage)
            if not site_records.empty and 'MaterialID' in site_records.columns:
                # MaterialID 기준으로 Usage 합산 및 Name/Spec 조회
                mat_groups = {}  # key: MaterialID_upper -> {'qty': float, 'name': str, 'spec': str, 'original_id': str}
                for _, row in site_records.iterrows():
                    mat_id = str(row.get('MaterialID', '')).strip()
                    if not mat_id or mat_id == 'nan':
                        continue
                    qty = 0
                    try: qty = float(row.get('Usage', 0) or 0)
                    except: pass

                    mat_id_upper = mat_id.upper()
                    if mat_id_upper not in mat_groups:
                        # MaterialID를 사용하여 materials_df에서 실제 품목명과 규격 조회
                        mat_name_val = mat_id
                        mat_spec_val = ''
                        disp_name = ''
                        mat_cat_val = ''
                        
                        if hasattr(self, 'materials_df') and not self.materials_df.empty:
                            try:
                                # 숫자형 ID 대응 (ID가 405.0 등일 수 있음)
                                clean_id = str(mat_id).strip().replace('.0', '')
                                match = self.materials_df[
                                    self.materials_df['MaterialID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True) == clean_id
                                ]
                                if not match.empty:
                                    # DB에서 값을 가져오되, 품목명 안에 규격이 합쳐져 있을 경우를 대비해 다시 분리
                                    raw_name = str(match.iloc[0].get('품목명', mat_name_val)).strip()
                                    raw_spec = str(match.iloc[0].get('규격', '')).strip()
                                    
                                    # [USER REQUEST] Keep 'Carestream' prefix
                                    # (Removal logic removed)
                                    
                                    if '-' in raw_name and (not raw_spec or raw_spec == 'nan' or raw_spec == '' or raw_spec == '자동등록'):
                                        # 품목명에 하이픈이 있고 규격 칸이 비어있거나 무의미한 경우 분리 실행
                                        dash_idx = raw_name.index('-')
                                        mat_name_val = raw_name[:dash_idx].strip()
                                        mat_spec_val = raw_name[dash_idx+1:].strip()
                                    else:
                                        mat_name_val = raw_name
                                        mat_spec_val = raw_spec if (raw_spec and str(raw_spec).lower() != 'nan' and raw_spec != '자동등록') else ''
                                    
                                    # [NEW] Get Category (Classification)
                                    mat_cat_val = str(match.iloc[0].get('품목군코드', '')).strip()
                                    if not mat_cat_val: mat_cat_val = str(match.iloc[0].iloc[8]) if len(match.columns) > 8 else ''
                            except: pass
                        
                        # [FIX] If we still only have an ID (like '410') and it contains a dash, try splitting it
                        if mat_name_val == mat_id and '-' in mat_id:
                            dash_idx = mat_id.index('-')
                            mat_name_val = mat_id[:dash_idx].strip()
                            mat_spec_val = mat_id[dash_idx+1:].strip()

                        disp_name = f"{mat_name_val}-{mat_spec_val}".upper() if mat_spec_val else mat_name_val.upper()
                        
                        mat_groups[mat_id_upper] = {
                            'qty': 0, 'name': mat_name_val, 'spec': mat_spec_val, 'original_id': mat_id, 'disp': disp_name,
                            'category': mat_cat_val
                        }
                    mat_groups[mat_id_upper]['qty'] += qty

                # [REFINEMENT] Secondary Grouping by Display Name to merge identical items
                merged_mats = {} # key: disp_name -> grp_data
                for _, grp in mat_groups.items():
                    d_name = grp.get('disp', 'Unknown')
                    if d_name not in merged_mats:
                        merged_mats[d_name] = grp.copy()
                    else:
                        merged_mats[d_name]['qty'] += grp['qty']

                # Build data['materials'] using merged results
                rt_keys = ['RT T200', 'RT AA400', 'RT Other']
                rt_counter = 0
                
                # [DEBUG] Log all merged materials before filtering
                print(f"DEBUG: All merged materials before RT filtering: {list(merged_mats.keys())}")
                
                # Separate RT items for sequential keying
                # [FIX] Include MT/PT keywords to prevent chemicals from being categorized as RT
                chem_names = [
                    'WHITE', 'BLACK', '7C', 'PENETRANT', 'CLEANER', 'DEVELOPER',
                    '백색', '흑색', '자분', '침투', '세척', '현상', '페인트',
                    'MT', 'PT'
                ]
                rt_merged = {}
                for d_name, grp in merged_mats.items():
                    is_chem = any(c in d_name.upper() for c in chem_names)
                    # [IMPROVED] Expanded RT detection to catch films by brand and model
                    rt_keywords = ['RT ', 'RT-', 'FILM', 'CARESTREAM', 'AGFA', 'FUJI', 'KODAK', 'T200', 'AA400', 'MX125', 'M100']
                    cat_upper = str(grp.get('category', '')).upper()
                    is_rt_cat = any(k in cat_upper for k in ['FILM', 'RT'])
                    is_rt_name = any(k in d_name.upper() for k in rt_keywords)
                    
                    if (is_rt_cat or is_rt_name) and not is_chem:
                        rt_merged[d_name] = grp

                for d_name, grp in rt_merged.items():
                    print(f"DEBUG: Processing RT Item {rt_counter+1}: {d_name}")
                    if rt_counter < 3:
                        mat_key = rt_keys[rt_counter]
                    else:
                        mat_key = f"RT_ROW_{rt_counter + 1}"
                    
                    rt_counter += 1
                    data['materials'][mat_key] = {
                        'used': float(grp['qty']), 
                        'name': grp['name'], 
                        'spec': grp['spec'], 
                        'is_rt': True, 
                        'category': 'RT'
                    }
                
                # [NOTE] Non-RT (Chemicals) are handled separately below

            else:
                # DB 데이터 없으면 UI에서 읽기 (fallback)
                if hasattr(self, 'ndt_company_entries') and self.ndt_company_entries:
                    mats = self.ndt_company_entries[0]
                    for m_key, mat_name in [('RT T200', 'T200'), ('RT AA400', 'AA400')]:
                        if m_key in mats:
                            val = mats[m_key].get().strip()
                            try: used = int(val) if val else 0
                            except: used = 0
                            data['materials'][m_key] = {'used': used}

            # [NEW] NDT 화학약품 (MT/PT) - DB site_records에서 합산하여 수집
            # [FIX] DT_ 접두어 포함하여 매칭 확장
            chem_db_map = [
                ('MT WHITE',     ['NDT_백색', 'NDT_백색페인트', 'NDT_형광자분', 'NDT_백색페인트_MT', 'DT_백색페인트', 'DT_형광자분']), 
                ('MT 7C-BLACK',  ['NDT_흑색', 'NDT_흑색자분', 'DT_흑색자분']),
                ('PT Penetrant', ['NDT_침투', 'NDT_침투제', 'DT_침투', 'DT_침투제']),
                ('PT Cleaner',   ['NDT_세척', 'NDT_세척제', 'DT_세척', 'DT_세척제']),
                ('PT Developer', ['NDT_현상', 'NDT_현상제', 'DT_현상', 'DT_현상제']),
            ]
            
            db_chem_found = False
            if not site_records.empty:
                # [FIX] DB 컬럼명에 공백이 있을 경우를 대비해 유연하게 매칭 (Normalization)
                actual_cols = list(site_records.columns)
                
                for m_key, db_cols in chem_db_map:
                    val_sum = 0
                    # [NEW] Check if any rows in site_records have a Material name matching this chemical
                    # and if so, sum up their 'Usage' or '수량' columns if they look like chemicals
                    for idx, row in site_records.iterrows():
                        m_name = str(row.get('Material', row.get('품목명', ''))).upper()
                        # Check if this row's material matches any of our chemical patterns
                        is_match = False
                        for col_pattern in db_cols:
                            pattern_norm = col_pattern.replace('NDT_', '').replace('DT_', '').replace(' ', '').upper()
                            if pattern_norm in m_name.replace(' ', ''):
                                is_match = True
                                break
                        
                        if is_match:
                            # If it matches, try 'Usage' first, then '수량' if it looks small (like cans)
                            u_val = pd.to_numeric(row.get('Usage'), errors='coerce')
                            if pd.isna(u_val): u_val = 0
                            val_sum += u_val
                            db_chem_found = True

                    # Also check specific columns if they exist as fallback/alternative
                    for col_pattern in db_cols:
                        pattern_norm = col_pattern.replace(' ', '').upper()
                        found_col = None
                        for actual_col in actual_cols:
                            if str(actual_col).replace(' ', '').upper() == pattern_norm:
                                found_col = actual_col
                                break
                        
                        if found_col:
                            try:
                                val_sum += int(pd.to_numeric(site_records[found_col], errors='coerce').fillna(0).sum())
                                db_chem_found = True
                            except: pass
                    
                    if val_sum > 0:
                        # [FIX] Use float for chemical quantities
                        data['materials'][m_key] = {'used': float(val_sum)}
                        db_chem_found = True
            
            # Debug gathered materials data
            print(f"DEBUG: Gathered Materials Data for Report: {data['materials']}")
            
            # DB에 데이터가 없으면 UI 위젯에서 읽기 (Fallback)
            if not db_chem_found and hasattr(self, 'ndt_company_entries') and self.ndt_company_entries:
                entries_dict = self.ndt_company_entries[0]
                mat_keys = [k for k in entries_dict.keys() if k != '_company']
                fallback_order = [
                    ('MT WHITE', 0), ('MT 7C-BLACK', 1), 
                    ('PT Penetrant', 3), ('PT Cleaner', 4), ('PT Developer', 5)
                ]
                for m_key, idx in fallback_order:
                    if idx < len(mat_keys):
                        try:
                            val = entries_dict[mat_keys[idx]].get().strip()
                            used = int(val) if val else 0
                            if used > 0: data['materials'][m_key] = {'used': used}
                        except: pass

            # 2.5 OT Status Gathering
            data['ot_status'] = []
            
            def clean_val(v):
                if pd.isna(v) or str(v).lower() == 'nan': return ""
                val = str(v).strip()
                # Strip common titles from both END and START
                titles = ["부장", "차장", "과장", "대리", "주임", "계장", "사원", "반장", "기사"]
                for title in titles:
                    if val.endswith(title): val = val[:-len(title)].strip()
                    if val.startswith(title): val = val[len(title):].strip()
                return val

            # 1) Priority: Check DB (Site Records) first for "Saved" data
            if not site_records.empty:
                print(f"DEBUG: --- ALL DB COLUMNS: {site_records.columns.tolist()} ---")
                for idx, row in site_records.iterrows():
                    # [FINAL FIX] Scan all 10 user columns to ensure no one is missed (e.g. User2, User3...)
                    for i in range(1, 11):
                        u_key = 'User' if i == 1 else f'User{i}'
                        wt_key = 'WorkTime' if i == 1 else f'WorkTime{i}'
                        oa_key = 'OT' if i == 1 else f'OT{i}'
                        
                        worker_name = clean_val(row.get(u_key, row.get('작업자' if i==1 else f'작업자{i}', '')))
                        if not worker_name or worker_name == "": continue
                        
                        ot_val = clean_val(row.get(wt_key, ''))
                        if not ot_val:
                            ot_val = clean_val(row.get('작업시간' if i==1 else f'작업시간{i}', ''))
                        
                        # Fallback for WorkTime: If specific UserX's WorkTime is missing, try general WorkTime column
                        if not ot_val:
                            ot_val = clean_val(row.get('WorkTime', row.get('작업시간', '')))
                        
                        if not ot_val:
                            # Scan all columns for time-like strings if explicit names fail
                            for col_name in row.index:
                                potential_val = str(row[col_name])
                                if any(k in potential_val for k in ["09:00", "24:00", "(주야간)", "~"]):
                                    ot_val = potential_val.strip()
                                    break
                        
                        if not ot_val:
                            ot_val = clean_val(row.get('OT시간' if i==1 else f'OT시간{i}', ''))
                            
                        # For amount, 'OT' seems to have it in this DB
                        ot_amount = clean_val(row.get(oa_key, row.get('OT금액' if i==1 else f'OT금액{i}', '')))
                        if not ot_amount and i > 1:
                            # Fallback to general OT column if specific one is empty
                            ot_amount = clean_val(row.get('OT', row.get('OT금액', '')))
                        
                        data['ot_status'].append({
                            'names': worker_name,
                            'company': clean_val(row.get('업체명', '')),
                            'method': clean_val(row.get('검사방법', '')),
                            'ot_hours': ot_val,
                            'ot_amount': ot_amount
                        })
                print(f"DEBUG: Gathered {len(data['ot_status'])} OT items from DB")

            # 2) Fallback: If DB is empty, check UI entries
            if not data['ot_status']:
                for i in range(1, 11):
                    group = getattr(self, f'worker_group{i}', None)
                    if group:
                        try:
                            name = group.get_worker().strip()
                            if name and name.lower() != 'nan':
                                ot_val = ""
                                if hasattr(group, 'ent_ot'): ot_val = group.ent_ot.get().strip()
                                elif hasattr(group, 'cb_ot'): ot_val = group.cb_ot.get().strip()
                                    
                                data['ot_status'].append({
                                    'names': name,
                                    'company': self.cb_daily_company.get().strip(),
                                    'method': self.cb_daily_test_method.get().strip(),
                                    'ot_hours': ot_val,
                                    'ot_amount': '' 
                                })
                        except: pass
                if data['ot_status']:
                    print(f"DEBUG: Gathered {len(data['ot_status'])} OT items from UI (Fallback)")
            
            print(f"DEBUG: FINAL OT Status for report: {len(data['ot_status'])} items")

            data['rtk'] = {}
            rtk_cats = {
                '센터미스': 'center_miss', '농도': 'density', '마킹미스': 'marking_miss',
                '필름마크': 'film_mark', '취급부주의': 'handling', '고객불만': 'customer_complaint', '기타': 'etc'
            }
            rtk_total = 0
            if not site_records.empty:
                # Get clean column names for matching
                import re as _re
                def clean_name(s): return _re.sub(r'[^A-Z가-힣0-9]', '', str(s).upper())
                
                actual_cols = site_records.columns.tolist()
                clean_cols = [clean_name(c) for c in actual_cols]
                
                for kor_key, eng_key in rtk_cats.items():
                    db_val = 0
                    target_clean = clean_name(kor_key)
                    rtk_target_clean = clean_name(f"RTK_{kor_key}")
                    
                    found_col = None
                    for i, c_clean in enumerate(clean_cols):
                        if c_clean == target_clean or c_clean == rtk_target_clean:
                            found_col = actual_cols[i]
                            break
                    
                    if found_col:
                        try:
                            # Sum up and handle potential string/NaN values
                            series = pd.to_numeric(site_records[found_col], errors='coerce').fillna(0)
                            db_val = int(series.sum())
                            print(f"DEBUG: Found RTK Col '{found_col}' for '{kor_key}', sum={db_val}")
                        except Exception as e:
                            print(f"DEBUG: Error summing RTK {kor_key}: {e}")
                    
                    data['rtk'][kor_key] = db_val
                    rtk_total += db_val
            
            # [CRITICAL FIX] If DB result is 0 or records empty, fallback to UI entries
            if rtk_total == 0:
                for kor_key, eng_key in rtk_cats.items():
                    val = 0
                    # Try direct widget attribute first
                    widget = getattr(self, f"ent_rtk_{eng_key}", None)
                    if widget:
                        try: val = int(widget.get().strip() or 0)
                        except: pass
                    # Then try the rtk_entries dictionary as backup
                    elif hasattr(self, 'rtk_entries') and kor_key in self.rtk_entries:
                        try: val = int(self.rtk_entries[kor_key].get().strip() or 0)
                        except: pass
                    
                    data['rtk'][kor_key] = val
                    rtk_total += val
                    
            data['rtk_total'] = rtk_total

            default_filename = f"작업일보_{site}_{date_val.strftime('%Y%m%d')}.xlsx"
            save_path = filedialog.asksaveasfilename(
                title="작업일보 저장",
                initialfile=default_filename,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )

            if not save_path:
                return

            # 4. 리포트 생성
            mapping = self.load_report_mapping()
            manager = DailyWorkReportManager(template_path)
            manager.generate_report(data, save_path, custom_mapping=mapping)

            messagebox.showinfo("성공", f"작업일보가 생성되었습니다:\n{save_path}")
            
            # 생성된 파일 열기
            if messagebox.askyesno("확인", "생성된 작업일보를 지금 여시겠습니까?"):
                os.startfile(save_path)

        except PermissionError:
            messagebox.showerror("오류", f"파일 접근 오류: '{os.path.basename(save_path)}' 파일이 다른 프로그램(엑셀 등)에서 열려 있어 저장할 수 없습니다.\n파일을 닫고 다시 시도해 주세요.")
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("오류", f"작업일보 생성 중 오류 발생: {e}")

    def _get_usage_session_data(self):
        """Helper to collect Site tab entry data"""
        data = {
            'date': str(self.ent_daily_date.get_date()),
            'site': self.cb_daily_site.get().strip(),
            'company': self.cb_daily_company.get().strip(),
            'applied_code': self.ent_daily_applied_code.get().strip(),
            'report_no': self.ent_daily_report_no.get().strip(),
            'equip': self.cb_daily_equip.get().strip(),
            'material': self.cb_daily_material.get().strip(),
            'inspection_item': self.ent_daily_inspection_item.get().strip(),
            'method': self.cb_daily_test_method.get().strip(),
            'unit': self.cb_daily_unit.get().strip(),
            'amount': self.ent_daily_test_amount.get().strip(),
            'unit_price': self.ent_daily_unit_price.get().strip(),
            'travel_cost': self.ent_daily_travel_cost.get().strip(),
            'meal_cost': self.ent_daily_meal_cost.get().strip(),
            'test_fee': self.ent_daily_test_fee.get().strip(),
            'note': self.ent_daily_note.get().strip(),
            'workers': []
        }
        for i in range(1, 11):
            group = getattr(self, f'worker_group{i}', None)
            if group:
                data['workers'].append({
                    'name': group.cb_name.get().strip(),
                    'shift': group.cb_shift.get().strip(),
                    'time': group.ent_worktime.get().strip(),
                    'ot': group.ent_ot.get().strip()
                })
        data['ndt_companies'] = []
        if hasattr(self, 'ndt_company_entries'):
            for entries in self.ndt_company_entries:
                co_data = {'_company': entries['_company'].get().strip()}
                for mat in ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]:
                    co_data[mat] = entries[mat].get().strip()
                data['ndt_companies'].append(co_data)
        data['rtk'] = {cat: ent.get().strip() for cat, ent in self.rtk_entries.items() if cat != '총계'}
        
        # [NEW] Include vehicle inspection data in session snapshot
        data['vehicles'] = []
        if hasattr(self, 'vehicle_boxes'):
            for box in self.vehicle_boxes:
                data['vehicles'].append(box.get_data())
                
        return data

    def _set_usage_session_data(self, data):
        """Helper to restore Site tab entry data"""
        if not data: return
        self.ent_daily_date.set_date(data.get('date', datetime.datetime.now().strftime('%Y-%m-%d')))
        # [FIX] Use the dedicated Company field
        if 'applied_code' in data:
            self.ent_daily_applied_code.delete(0, tk.END)
            self.ent_daily_applied_code.insert(0, data.get('applied_code', ''))
        
        self.cb_daily_company.set(data.get('company', ''))
        self.cb_daily_site.set(data.get('site', ''))
        self.ent_daily_report_no.delete(0, tk.END); self.ent_daily_report_no.insert(0, data.get('report_no', ''))
        if isinstance(self.cb_daily_material, ttk.Combobox):
            self.cb_daily_material.set(data.get('material', ''))
        else:
            self.cb_daily_material.delete(0, tk.END)
            self.cb_daily_material.insert(0, data.get('material', ''))
        if 'inspection_item' in data:
            self.ent_daily_inspection_item.delete(0, tk.END)
            self.ent_daily_inspection_item.insert(0, data.get('inspection_item', ''))
        self.cb_daily_test_method.set(data.get('method', ''))
        self.cb_daily_unit.set(data.get('unit', '매'))
        self.ent_daily_test_amount.delete(0, tk.END); self.ent_daily_test_amount.insert(0, data.get('amount', ''))
        self.ent_daily_unit_price.delete(0, tk.END); self.ent_daily_unit_price.insert(0, data.get('unit_price', ''))
        self.ent_daily_travel_cost.delete(0, tk.END); self.ent_daily_travel_cost.insert(0, data.get('travel_cost', ''))
        self.ent_daily_meal_cost.delete(0, tk.END); self.ent_daily_meal_cost.insert(0, data.get('meal_cost', ''))
        self.ent_daily_test_fee.delete(0, tk.END); self.ent_daily_test_fee.insert(0, data.get('test_fee', ''))
        self.ent_daily_note.delete(0, tk.END); self.ent_daily_note.insert(0, data.get('note', ''))
        workers = data.get('workers', [])
        for i, w_data in enumerate(workers, 1):
            group = getattr(self, f'worker_group{i}', None)
            if group:
                group.cb_name.set(w_data.get('name', ''))
                group.cb_shift.set(w_data.get('shift', ''))
                group.ent_worktime.set(w_data.get('time', ''))
                group.ent_ot.delete(0, tk.END); group.ent_ot.insert(0, w_data.get('ot', ''))
        ndt_data = data.get('ndt_companies', [])
        while len(self.ndt_company_entries) > 1: self.remove_last_ndt_company()
        for i, co_data in enumerate(ndt_data):
            if i >= len(self.ndt_company_entries): self.add_ndt_company_section()
            entries = self.ndt_company_entries[i]
            entries['_company'].set(co_data.get('_company', ''))
            for mat in ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]:
                entries[mat].delete(0, tk.END); entries[mat].insert(0, co_data.get(mat, ''))
        rtk = data.get('rtk', {})
        for cat, val in rtk.items():
            if cat in self.rtk_entries:
                ent = self.rtk_entries[cat]; ent.delete(0, tk.END); ent.insert(0, val)
        self.calculate_rtk_total()
        
        # [NEW] Restore vehicle inspection data
        vehicle_data = data.get('vehicles', [])
        if vehicle_data:
            # Match current boxes or add new ones if needed
            for i, v_state in enumerate(vehicle_data):
                if i < len(self.vehicle_boxes):
                    self.vehicle_boxes[i].set_data(v_state)
                else:
                    # If more boxes are needed than current, we could add them,
                    # but usually there's 1 or the count is already correct.
                    pass

    def _get_inout_session_data(self):
        """Helper to collect In/Out tab entry data"""
        data = {
            'reg': {
                'co_code': self.cb_co_code.get().strip(),
                'eq_code': self.cb_eq_code.get().strip(),
                'item_name': self.cb_item_name.get().strip(),
                'model': self.cb_model.get().strip(),
                'sn': self.ent_sn.get().strip(),
                'class': self.cb_class.get().strip(),
                'spec': self.cb_spec.get().strip(),
                'unit': self.cb_unit.get().strip(),
                'supplier': self.cb_supplier.get().strip(),
                'mfr': self.cb_mfr.get().strip(),
                'origin': self.cb_origin.get().strip(),
                'reorder': self.ent_reorder.get().strip(),
                'init': self.ent_init.get().strip(),
                'price': self.ent_price.get().strip(),
                'cost': self.ent_cost.get().strip()
            },
            'trans': {
                'material': self.cb_material.get().strip(),
                'type': self.cb_type.get().strip(),
                'qty': self.ent_qty.get().strip(),
                'site': self.cb_trans_site.get().strip(),
                'warehouse': self.cb_warehouse.get().strip(),
                'user': self.ent_user.get().strip(),
                'note': self.ent_note.get().strip()
            }
        }
        return data

    def _set_inout_session_data(self, data):
        """Helper to restore In/Out tab entry data"""
        if not data: return
        reg = data.get('reg', {})
        self.cb_co_code.set(reg.get('co_code', ''))
        self.cb_eq_code.set(reg.get('eq_code', ''))
        self.cb_item_name.set(reg.get('item_name', ''))
        self.cb_model.set(reg.get('model', ''))
        self.ent_sn.delete(0, tk.END); self.ent_sn.insert(0, reg.get('sn', ''))
        self.cb_class.set(reg.get('class', ''))
        self.cb_spec.set(reg.get('spec', ''))
        self.cb_unit.set(reg.get('unit', ''))
        self.cb_supplier.set(reg.get('supplier', ''))
        self.cb_mfr.set(reg.get('mfr', ''))
        self.cb_origin.set(reg.get('origin', ''))
        self.ent_reorder.delete(0, tk.END); self.ent_reorder.insert(0, reg.get('reorder', '0'))
        self.ent_init.delete(0, tk.END); self.ent_init.insert(0, reg.get('init', '0'))
        self.ent_price.delete(0, tk.END); self.ent_price.insert(0, reg.get('price', '0'))
        self.ent_cost.delete(0, tk.END); self.ent_cost.insert(0, reg.get('cost', '0'))
        
        trans = data.get('trans', {})
        self.cb_material.set(trans.get('material', ''))
        self.cb_type.set(trans.get('type', 'OUT'))
        self.ent_qty.delete(0, tk.END); self.ent_qty.insert(0, trans.get('qty', ''))
        self.cb_trans_site.set(trans.get('site', ''))
        self.cb_warehouse.set(trans.get('warehouse', ''))
        self.ent_user.set(trans.get('user', ''))
        self.ent_note.delete(0, tk.END); self.ent_note.insert(0, trans.get('note', ''))

    def _get_budget_session_data(self):
        """Helper to collect Budget tab entry data"""
        data = {
            'main': {k: w.get().strip() for k, w in self.budget_widgets.items()},
            'labor_detail': self.labor_detail_widget.get_data(),
            'material_detail': self.material_detail_widget.get_data(),
            'expense_detail': self.expense_detail_widget.get_data()
        }
        return data

    def _set_budget_session_data(self, data):
        """Helper to restore Budget tab entry data"""
        if not data: return
        main = data.get('main', {})
        for k, val in main.items():
            if k in self.budget_widgets:
                w = self.budget_widgets[k]
                if hasattr(w, 'set'): w.set(val)
                else: w.delete(0, tk.END); w.insert(0, val)
        self.labor_detail_widget.set_data(data.get('labor_detail', {}))
        self.material_detail_widget.set_data(data.get('material_detail', {}))
        self.expense_detail_widget.set_data(data.get('expense_detail', {}))
        self._update_budget_kpis()

    def save_form_session(self):
        """Orchestrate saving session data from all entry tabs into one JSON"""
        try:
            global_data = {
                'usage': self._get_usage_session_data(),
                'inout': self._get_inout_session_data(),
                'budget': self._get_budget_session_data(),
                'saved_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            file_path = filedialog.asksaveasfilename(
                title="전체 세션 저장 (일괄)",
                defaultextension=".json",
                filetypes=[("JSON files", "*.json")],
                initialfile=f"PMI_Global_Session_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            )
            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(global_data, f, ensure_ascii=False, indent=2)
                messagebox.showinfo("저장 완료", f"현장/입출고/실행예산 탭의 입력 내용이 일괄 저장되었습니다.\n{os.path.basename(file_path)}")
        except Exception as e:
            messagebox.showerror("오류", f"전체 세션 저장 중 오류 발생: {e}")

    def load_form_session(self):
        """Orchestrate loading session data back to all entry tabs from a JSON file"""
        try:
            file_path = filedialog.askopenfilename(
                title="전체 세션 불러오기 (일괄)",
                filetypes=[("JSON files", "*.json")]
            )
            if not file_path: return
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            # Restore all tabs
            self._set_usage_session_data(data.get('usage'))
            self._set_inout_session_data(data.get('inout'))
            self._set_budget_session_data(data.get('budget'))
            messagebox.showinfo("불러오기 완료", "현장/입출고/실행예산 탭의 데이터가 일괄 복원되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"전체 세션 불러오기 중 오류 발생: {e}")

    def _auto_reconcile_and_register_ndt(self, date_val, site, ndt_data, workers, company_code=""):
        """Unified helper to auto-register missing consumables and create 'OUT' transactions"""
        ndt_product_map = self._load_ndt_product_map()
        ndt_groups = self.ndt_groups
        
        note_suffix = f" ({company_code})" if company_code else ""
        note = f"{site} 현장 사용 (자동 차감){note_suffix}"
        
        # Ensure date_val is datetime and handle date vs datetime objects
        if isinstance(date_val, str):
            date_obj = pd.to_datetime(date_val)
        else:
            date_obj = date_val
            
        # extract date part safely
        pure_date = date_obj.date() if hasattr(date_obj, 'date') else date_obj
        trans_date = datetime.datetime.combine(pure_date, datetime.datetime.now().time())
        
        for name, qty in ndt_data.items():
            if qty <= 0: continue
            
            # 1. Resolve Item Name (Group) and Model
            db_item_name = ndt_product_map.get(name, '')
            db_model_name = name
            if not db_item_name:
                for grp, members in ndt_groups.items():
                    if name in members:
                        db_item_name = grp
                        db_model_name = name # [FIX] Remove prefix for consistency with display logic
                        break
                if not db_item_name: db_item_name = "기타소모품"
            
            # 2. Find or Register (Highly flexible matching for NDT)
            target_item = db_item_name.replace(' ', '')
            
            # [FIX] Fill NaN with empty string before comparison to ensure reliable matching
            m_df = self.materials_df.copy()
            for col in ['품목명', '모델명']:
                if col in m_df.columns:
                    m_df[col] = m_df[col].fillna('').astype(str).str.replace(' ', '', regex=False)

            # Try to find existing item
            # First, check for exact (item+model) match
            target_model = db_model_name.replace(' ', '')
            exact_match = m_df[(m_df['품목명'] == target_item) & (m_df['모델명'] == target_model)]
            
            if not exact_match.empty:
                mat_id = exact_match.iloc[0]['MaterialID']
            else:
                # If no exact match, try matching by item name ONLY for NDT groups
                item_only_match = m_df[m_df['품목명'] == target_item]
                if not item_only_match.empty:
                    # Use the first available item in this group (e.g., generic "PT약품")
                    mat_id = item_only_match.iloc[0]['MaterialID']
                else:
                    # Truly missing - [REFINED] Do NOT register new one automatically
                    print(f"DEBUG: Skipping auto-registration for missing NDT item '{db_item_name}' (Model: {db_model_name})")
                    mat_id = None
            
            # 3. Create Transaction (Only if mat_id exists)
            if mat_id:
                self._create_manual_stock_transaction(trans_date, mat_id, 'OUT', qty, site, workers, note)

    def _create_manual_stock_transaction(self, date_val, mat_id, trans_type, qty, site, user, note):
        # [V23.1_WATERTIGHT_FILTER] Default to block everything unless it matches the whitelist
        try:
            if self.materials_df.empty:
                return # Can't identify item, don't save
                
            # Robust ID matching
            str_target_id = str(mat_id).strip().replace('.0', '')
            # Vectorized lookup for speed
            mask = self.materials_df['MaterialID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True) == str_target_id
            mat_info = self.materials_df[mask]
            
            if mat_info.empty:
                # If we can't find it in DB, we don't know what it is. To be safe, DON'T save it to history.
                return
                
            row = mat_info.iloc[0]
            item_name = str(row.get('품목명', '')).strip()
            model_name = str(row.get('모델명', '')).strip()
            item_name_up = item_name.upper()
            model_name_up = model_name.upper()
            
            # 1. Check for Carestream (Partial Match, Case Insensitive)
            is_carestream = "CARESTREAM" in item_name_up or "CARESTREAM" in model_name_up
            
            # 2. Check for NDT Chemicals (More inclusive partial match for PT/MT/NDT)
            ndt_keywords = ["PT", "MT", "NDT", "침투제", "세척제", "현상제", "자분", "페인트"]
            is_ndt_chem = any(kw in item_name_up or kw in model_name_up for kw in ndt_keywords)
            
            # [CRITICAL] If not a whitelist item, exit NOW
            if not (is_carestream or is_ndt_chem):
                return
                
        except Exception as e:
            print(f"Filter error in _create_manual_stock_transaction: {e}")
            return # Block on error
            
        # --- IF WE GOT HERE, THE ITEM IS ALLOWED ---
        try:
            # Ensure quantity is negative for OUT
            signed_qty = -abs(float(qty)) if trans_type == 'OUT' else abs(float(qty))
            
            new_trans = {
                'Date': pd.to_datetime(date_val),
                'MaterialID': mat_id,
                'Type': trans_type,
                'Quantity': signed_qty,
                'User': user,
                'Site': site,
                'Note': note
            }
            self.transactions_df = pd.concat([self.transactions_df, pd.DataFrame([new_trans])], ignore_index=True)
        except: pass
    
    def refresh_inquiry_filters(self):
        """Unified method to populate inquiry filter dropdowns and sync with autocomplete lists"""
        try:
            import re

            marker_pattern = MARKER_PATTERN
            
            # 1. Collect and Merge Site Info
            raw_sites = set()
            if not self.daily_usage_df.empty and 'Site' in self.daily_usage_df.columns:
                raw_sites.update(self.daily_usage_df['Site'].dropna().astype(str).apply(self.normalize_site_name).tolist())
            if hasattr(self, 'sites'):
                raw_sites.update(self.sites)
            # [NEW] budget_df에 저장된 현장명도 포함
            if hasattr(self, 'budget_df') and not self.budget_df.empty and 'Site' in self.budget_df.columns:
                raw_sites.update(self.budget_df['Site'].dropna().astype(str).apply(self.normalize_site_name).tolist())
            unique_sites = sorted([s for s in raw_sites 
                                     if s and str(s).lower() != 'nan'
                                    and s not in getattr(self, 'hidden_sites', [])])  # [FIX] 숨긴 현장 제외
            
            # [ROBUST] Do NOT overwrite self.sites Master List with history.
            # self.sites is the "Managed List". unique_sites is for "Filters".
            # self.sites[:] = unique_sites # REMOVED: This was causing deleted sites to reappear
            
            if hasattr(self, 'cb_daily_filter_site'):
                self.cb_daily_filter_site['values'] = ['전체'] + unique_sites
                if not self.cb_daily_filter_site.get(): self.cb_daily_filter_site.set('전체')
            if hasattr(self, 'cb_filter_site_monthly'):
                self.cb_filter_site_monthly['values'] = ['전체'] + unique_sites
            # [FIX] 공사실행예산서 상단 현장 콤보박스 → budget_df 현장만 표시
            # (삭제 후 목록에서 즉시 사라지게 하기 위해 daily_usage_df 현장은 제외)
            if hasattr(self, 'cb_budget_site'):
                budget_sites = []
                if hasattr(self, 'budget_df') and not self.budget_df.empty and 'Site' in self.budget_df.columns:
                    budget_sites = sorted([
                        s for s in self.budget_df['Site'].dropna().astype(str).apply(self.normalize_site_name).unique()
                        if s and str(s).lower() != 'nan'
                    ])
                self.budget_sites[:] = budget_sites  # [FIX] 자동완성 목록 동기화
                current_val = self.cb_budget_site.get()
                self.cb_budget_site['values'] = budget_sites
                if current_val: self.cb_budget_site.set(current_val)
            # 하단 실적 조회 콤보박스 → 모든 현장(daily_usage_df + budget_df) 표시
            if hasattr(self, 'cb_budget_view_site'):
                self.cb_budget_view_site['values'] = unique_sites

            # 2. Collect Material Info (History Only)
            raw_materials = set()
            if not self.daily_usage_df.empty and 'MaterialID' in self.daily_usage_df.columns:
                unique_mat_ids = self.daily_usage_df['MaterialID'].dropna().unique()
                for mat_id in unique_mat_ids:
                    name = self.get_material_display_name(mat_id)
                    if name: raw_materials.add(name)
            
            unique_materials = sorted(list(raw_materials))
            
            if hasattr(self, 'cb_daily_filter_material'):
                # [RESTORED] Populate with history names to avoid 'empty' confusion, while still excluding master list
                self.cb_daily_filter_material['values'] = ['전체'] + unique_materials
                if not self.cb_daily_filter_material.get(): self.cb_daily_filter_material.set('전체')
            
            if hasattr(self, 'cb_daily_material'):
                # [NEW] Only update values if it's still a combobox (Site tab uses Entry now)
                if isinstance(self.cb_daily_material, ttk.Combobox):
                    self.cb_daily_material['values'] = unique_materials
            
            if hasattr(self, 'cb_daily_filter_site'):
                self.cb_daily_filter_site['values'] = ['전체'] + unique_sites
            if hasattr(self, 'cb_filter_material_monthly'):
                self.cb_filter_material_monthly['values'] = ['전체'] + unique_materials

            # 3. Equipment dropdown
            if hasattr(self, 'cb_daily_filter_equipment'):
                raw_equip = []
                if not self.daily_usage_df.empty and '장비명' in self.daily_usage_df.columns:
                    raw_equip = self.daily_usage_df['장비명'].dropna().astype(str).str.strip().unique().tolist()
                
                # Combine with catalog if needed
                unique_equip = sorted(list(set([e for e in raw_equip if e and str(e).lower() != 'nan'])))
                self.cb_daily_filter_equipment['values'] = ['전체'] + unique_equip
                if not self.cb_daily_filter_equipment.get(): self.cb_daily_filter_equipment.set('전체')

            # 4. Worker dropdown
            if hasattr(self, 'cb_daily_filter_worker'):
                worker_cols = ['User', 'User2', 'User3', 'User4', 'User5', 'User6', 'User7', 'User8', 'User9', 'User10']
                all_workers_raw = set()
                if not self.daily_usage_df.empty:
                    for col in worker_cols:
                        if col in self.daily_usage_df.columns:
                            all_workers_raw.update(self.daily_usage_df[col].dropna().astype(str).unique().tolist())
                
                # Include static list
                if hasattr(self, 'users'): all_workers_raw.update(self.users)
                
                clean_workers_set = set()
                for w in all_workers_raw:
                    w_str = str(w).strip()
                    if not w_str or w_str.lower() in ['nan', '0.0', 'none']: continue
                    cleaned = marker_pattern.sub('', w_str).strip()
                    cleaned = " ".join(cleaned.split())
                    if cleaned: clean_workers_set.add(cleaned)

                unique_workers = sorted(list(clean_workers_set))
                # self.users[:] = unique_workers # REMOVED: This was causing deleted workers to reappear
                self.cb_daily_filter_worker['values'] = ['전체'] + unique_workers
                if not self.cb_daily_filter_worker.get(): self.cb_daily_filter_worker.set('전체')

            # 5. Vehicle dropdown
            raw_vehicles = set()
            if not self.daily_usage_df.empty and '차량번호' in self.daily_usage_df.columns:
                v_list = self.daily_usage_df['차량번호'].dropna().astype(str).str.strip().tolist()
                for v in v_list:
                    cleaned = self.clean_nan(v)
                    if cleaned: raw_vehicles.add(cleaned)
            if hasattr(self, 'vehicles'):
                raw_vehicles.update(self.vehicles)
                
            unique_vehicles = sorted(list([v for v in raw_vehicles if v and str(v).lower() != 'nan']))
            if hasattr(self, 'cb_daily_filter_vehicle'):
                self.cb_daily_filter_vehicle['values'] = ['전체'] + unique_vehicles
                if not self.cb_daily_filter_vehicle.get(): self.cb_daily_filter_vehicle.set('전체')
                
            # 6. Company dropdown (for entry form)
            raw_companies = set()
            if not self.daily_usage_df.empty and '업체명' in self.daily_usage_df.columns:
                raw_companies.update(self.daily_usage_df['업체명'].dropna().astype(str).str.strip().tolist())
            if hasattr(self, 'companies'):
                raw_companies.update(self.companies)
                
            unique_companies = sorted(list([c for c in raw_companies if c and str(c).lower() != 'nan']))
            if hasattr(self, 'cb_daily_company'):
                self.cb_daily_company['values'] = unique_companies
            if hasattr(self, 'cb_daily_filter_company'):
                self.cb_daily_filter_company['values'] = ['전체'] + unique_companies
                if not self.cb_daily_filter_company.get(): self.cb_daily_filter_company.set('전체')
            if hasattr(self, 'cb_daily_filter_vehicle'):
                self.cb_daily_filter_vehicle['values'] = ['전체'] + unique_vehicles
                if not self.cb_daily_filter_vehicle.get(): self.cb_daily_filter_vehicle.set('전체')
        except Exception as e:
            print(f"ERROR in refresh_inquiry_filters: {e}")
        if hasattr(self, 'cb_sales_filter_site'):
            self.cb_sales_filter_site['values'] = ['전체'] + self.sites
            if not self.cb_sales_filter_site.get(): self.cb_sales_filter_site.set('전체')
        if hasattr(self, 'cb_trans_filter_vehicle'):
            self.cb_trans_filter_vehicle['values'] = ['전체'] + unique_vehicles

    def update_daily_usage_view(self):
        """Update the daily usage treeview with smarter filters and shift classification"""
        import re
        marker_pattern = MARKER_PATTERN
        
        # [FIX] Aggressively hide any active suggestion window
        filter_widgets = [
            getattr(self, 'cb_daily_filter_site', None),
            getattr(self, 'cb_daily_filter_material', None),
            getattr(self, 'cb_daily_filter_equipment', None),
            getattr(self, 'cb_daily_filter_worker', None),
            getattr(self, 'cb_daily_filter_vehicle', None)
        ]
        for widget in filter_widgets:
            if widget and hasattr(widget, '_suggestion_win'):
                widget._suggestion_win.hide()

        # [REMOVED] focus_set here stoles focus from the entry form after saving
        # if hasattr(self, 'daily_usage_tree'):
        #     self.daily_usage_tree.focus_set()

        # Clear current view
        for item in self.daily_usage_tree.get_children():
            self.daily_usage_tree.delete(item)
        
        # Get filter values
        start_date_str = self.ent_daily_start_date.get().strip()
        end_date_str = self.ent_daily_end_date.get().strip()
        filter_site = self.cb_daily_filter_site.get().strip() if hasattr(self, 'cb_daily_filter_site') else '전체'
        filter_company = self.cb_daily_filter_company.get().strip() if hasattr(self, 'cb_daily_filter_company') else '전체'
        filter_material = self.cb_daily_filter_material.get().strip() if hasattr(self, 'cb_daily_filter_material') else '전체'
        filter_equipment = self.cb_daily_filter_equipment.get().strip() if hasattr(self, 'cb_daily_filter_equipment') else '전체'
        filter_worker = self.cb_daily_filter_worker.get().strip() if hasattr(self, 'cb_daily_filter_worker') else '전체'
        filter_vehicle = self.cb_daily_filter_vehicle.get().strip() if hasattr(self, 'cb_daily_filter_vehicle') else '전체'
        filter_shift = self.cb_daily_filter_shift.get().strip() if hasattr(self, 'cb_daily_filter_shift') else '전체'
        
        # Ensure default values if empty
        if not filter_site: filter_site = '전체'
        if not filter_material: filter_material = '전체'
        if not filter_equipment: filter_equipment = '전체'
        if not filter_worker: filter_worker = '전체'
        if not filter_shift: filter_shift = '전체'
        
        # (Dropdown population moved to refresh_inquiry_filters for stability)
        
        # [V20_FORCE_FIRST_LOAD] Ensure 2025 records are seen by resetting start date on first load
        if not hasattr(self, '_daily_usage_first_load_done'):
            self._daily_usage_first_load_done = True
            if hasattr(self, 'ent_daily_start_date'):
                # Force to 2024 to catch 2025-04 records
                self.ent_daily_start_date.set_date(datetime.datetime(2024, 1, 1))
                start_date_str = "2024-01-01"
                
        # Parse dates
        try:
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None
            if end_date:
                # Include the entire end date
                end_date = end_date + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)
        except ValueError:
            # Silently handle or log if needed, avoid blocking the entire view
            start_date = datetime.datetime(2024, 1, 1)
            end_date = datetime.datetime.now()
        
        # Filter data
        filtered_df = self.daily_usage_df.copy()
        # [CRITICAL] Normalize columns to ensure detection (site_pairs) matches data lookups
        filtered_df.columns = [str(c).strip().replace(' ', '') for c in filtered_df.columns]
        
        print(f"DEBUG: [Daily Usage] Total records in DB: {len(filtered_df)}")
        
        # [V17_CRASH_PROOF_DATE] Maximum resilience to prevent Exit Code 1
        def robust_date_parse(val):
            if val is None or pd.isna(val) or str(val).strip() == '': return pd.NaT
            try:
                # 1. Handle already parsed dates
                if isinstance(val, (pd.Timestamp, datetime.datetime)):
                    return val.replace(tzinfo=None)
                if isinstance(val, datetime.date):
                    return pd.Timestamp(val)
                    
                # 2. Handle Numeric/Excel Serial
                if isinstance(val, (int, float)):
                    if 30000 < val < 60000:
                        return pd.to_datetime(val, unit='D', origin='1899-12-30').round('min').replace(tzinfo=None)
                
                # 3. Handle Strings
                s_val = str(val).strip()
                if s_val.replace('.','').isdigit():
                    num = float(s_val)
                    if 30000 < num < 60000:
                        return pd.to_datetime(num, unit='D', origin='1899-12-30').round('min').replace(tzinfo=None)
                
                # Standard pandas parse with dot-to-hyphen cleanup
                clean_val = s_val.replace('.', '-').replace(' ', '')
                d = pd.to_datetime(clean_val, errors='coerce')
                if pd.notna(d): return d.replace(tzinfo=None)
                
                # Final raw parse attempt
                d = pd.to_datetime(s_val, errors='coerce')
                if pd.notna(d): return d.replace(tzinfo=None)
            except:
                pass
            return pd.NaT

        # Apply robust parsing
        filtered_df['Date'] = filtered_df['Date'].apply(robust_date_parse)
        
        # [V14_REVERTED_STILL] Sort by Date (Descending)
        sort_date_col = 'Date'
        if 'EntryTime' in filtered_df.columns:
            filtered_df['EntryTime'] = filtered_df['EntryTime'].apply(robust_date_parse)
            # Use EntryTime as tie-breaker (Newest to Oldest)
            filtered_df = filtered_df.sort_values(by=['Date', 'EntryTime'], ascending=[False, False], na_position='last')
        else:
            filtered_df = filtered_df.sort_values(by=['Date'], ascending=[False], na_position='last')
            
        # [V16_INDEX_PRESERVE] Removed reset_index to keep original mapping to self.daily_usage_df
        # filtered_df = filtered_df.reset_index(drop=True)
        
        if start_date is not None:
            filtered_df = filtered_df[filtered_df['Date'] >= start_date]
        
        if end_date is not None:
            filtered_df = filtered_df[filtered_df['Date'] <= end_date]
        
        if filter_site != '전체' and 'Site' in filtered_df.columns:
            # [V18_LOOSE_FILTER] Use partial match for site to avoid missing records with minor spacing/naming differences
            filtered_df = filtered_df[filtered_df['Site'].astype(str).str.contains(filter_site, case=False, na=False, regex=False)]
        
        if filter_company != '전체' and '업체명' in filtered_df.columns:
            filtered_df = filtered_df[filtered_df['업체명'].astype(str).str.contains(filter_company, case=False, na=False, regex=False)]
            
        if filter_material != '전체' and 'MaterialID' in filtered_df.columns:
            # Also allow partial match on display name
            def check_mat(mid):
                d_name = self.get_material_display_name(mid)
                return filter_material.lower() in d_name.lower()
            filtered_df = filtered_df[filtered_df['MaterialID'].apply(check_mat)]
            
        if filter_worker != '전체':
            # Check all 10 worker columns
            worker_cols = ['User'] + [f'User{i}' for i in range(2, 11)]
            mask = pd.Series([False] * len(filtered_df), index=filtered_df.index)
            for col in worker_cols:
                if col in filtered_df.columns:
                    mask |= filtered_df[col].astype(str).str.contains(filter_worker, case=False, na=False, regex=False)
            filtered_df = filtered_df[mask]
            # Support partial match for equipment
            filtered_df = filtered_df[filtered_df['장비명'].astype(str).str.contains(filter_equipment, na=False, case=False, regex=False)]
        
        if filter_vehicle != '전체' and '차량번호' in filtered_df.columns:
            # Support partial match for vehicle number
            filtered_df = filtered_df[filtered_df['차량번호'].astype(str).str.contains(filter_vehicle, na=False, case=False, regex=False)]
        
        # Filter by material if specified
        if filter_material != '전체':
            # Robust filtering: map every MaterialID in the current set to its display name 
            # and filter by matching the user's selection string. 
            # This correctly handles both master-linked and orphaned/deleted materials.
            def get_disp_name(mid): return self.get_material_display_name(mid)
            filtered_df = filtered_df[filtered_df['MaterialID'].apply(get_disp_name) == filter_material]

        # Filter by worker or shift (Smarter & Marker-Insensitive)
        if filter_worker != '전체' or filter_shift != '전체':
            worker_cols = ['User', 'User2', 'User3', 'User4', 'User5', 'User6', 'User7', 'User8', 'User9', 'User10']
            time_cols = ['WorkTime', 'WorkTime2', 'WorkTime3', 'WorkTime4', 'WorkTime5', 'WorkTime6', 'WorkTime7', 'WorkTime8', 'WorkTime9', 'WorkTime10']
            
            # Normalize filter text: remove spaces and lowercase for maximum flexibility
            fw_clean = filter_worker.replace(' ', '').lower()
            if fw_clean == '': fw_clean = '전체' # Handle empty input as showing all

            def row_matches(row):
                if fw_clean == '전체' and filter_shift == '전체':
                    return True
                
                # Pre-clean filter text for comparison
                f_worker = fw_clean
                f_shift = f"({filter_shift})" if filter_shift != '전체' else None
                
                for i in range(len(worker_cols)):
                    # Securely get worker and time data
                    w_col = worker_cols[i]
                    t_col = time_cols[i]
                    
                    if w_col not in row: continue
                    w_val_raw = str(row[w_col]).strip()
                    
                    if not w_val_raw or w_val_raw.lower() in ['nan', '0.0', 'none', '']:
                        continue
                        
                    t_val = str(row.get(t_col, '')).strip()
                    
                    # Worker match (Substring match on cleaned names)
                    w_match = True
                    if f_worker != '전체':
                        w_val_clean = w_val_raw.replace(' ', '').lower()
                        # Also handle records where marker is still in User column
                        w_val_clean = marker_pattern.sub('', w_val_clean).strip()
                        w_match = f_worker in w_val_clean
                    
                    # Shift match (Look for marker in WorkTime column)
                    s_match = True
                    if f_shift:
                        s_match = f_shift in t_val
                    
                    if w_match and s_match:
                        return True
                return False

            # Apply robust filtering
            filtered_df = filtered_df[filtered_df.apply(row_matches, axis=1)]
        
        # [V10] Sorting is now handled at the beginning of the function for better reliability
        
        # Define RTK categories
        rtk_categories = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타", "총계"]
        
        # Display entries and calculate totals
        total_rtk = [0.0] * len(rtk_categories)
        total_ndt = [0.0] * 7
        total_test_amount = 0.0
        total_unit_price = 0.0
        total_travel_cost = 0.0
        total_meal_cost = 0.0
        total_test_fee = 0.0
        total_film_count = 0.0 # [NEW] Added for film count
        total_ot_hours = 0.0
        total_ot_amount = 0
        total_work_hours = 0.0 # Added for Total Working Time
        total_mileage = 0.0    # Sum of individual entries if applicable
        min_mileage = float('inf')
        max_mileage = float('-inf')
        total_indiv_ot_hours = [0.0] * 10
        total_indiv_ot_amounts = [0] * 10
        
        has_equip = False
        has_method = False
        has_note = False
        has_product = False
        has_entry_time = False
        has_vehicle_no = False
        has_mileage = False
        has_veh_insp = False
        has_veh_note = False
        has_co_code = False
        has_company = False
        has_applied_code = False # [NEW]
        has_insp_item = False    # [NEW]
        has_report_no = False    # [NEW]
        has_meal_cost = False    # [NEW]
        has_unit = False         # [NEW]
        
        
        current_date = None
        
        # Seen sets for deduping
        seen_entry_times = set()
        seen_contents = set()
        
        for idx, row in filtered_df.iterrows():
            entry = row.to_dict()
            
            # Metadata formatting
            usage_date = self._safe_format_datetime(entry.get('Date', ''), '%Y-%m-%d')
            if not usage_date: usage_date = "Unknown"
            
            mat_id = entry.get('MaterialID', '')
            mat_name = self.get_material_display_name(mat_id)
            
            # Re-calculate workers/worktime early for deduping
            import re as _re
            def clean_s(v): return self.clean_nan(v)
            raw_workers = []
            for j in range(1, 11):
                u_k = 'User' if j == 1 else f'User{j}'
                u_v = clean_s(entry.get(u_k, ''))
                if u_v: raw_workers.append(u_v)
            
            c_workers = self.format_worker_summary(raw_workers)
            c_worktime = clean_s(entry.get('WorkTime', '')) if raw_workers else ""

            # Timestamp-based Key (Safety) - Define early for use in deduplication keys
            e_t_raw = entry.get('EntryTime', '')
            try:
                if isinstance(e_t_raw, (pd.Timestamp, datetime.datetime)):
                    t_key = e_t_raw.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    t_key = str(e_t_raw).split('.')[0].strip()
            except:
                t_key = str(e_t_raw).strip()

            # Content-based Unique Key (Date, Site, WorkTime, Method, EntryTime, ItemInfo)
            # [FIX] Include EntryTime (t_key) to ensure separate saves are never merged visually.
            # Also include Material, Item, and Code for maximum granularity within a single save if needed.
            n_site = str(entry.get('Site', '')).strip()
            n_date = usage_date
            n_method = str(entry.get('검사방법', '')).strip()
            n_insp = str(entry.get('검사품명', '')).strip()
            n_code = str(entry.get('적용코드', '')).strip()
            
            content_key = (n_date, n_site, c_worktime, n_method, t_key, mat_id, n_insp, n_code)
            
            # Record is duplicate split IF (Timestamp matches) AND (Content matches)
            # [FIX] Changed to OR but because t_key is in content_key, separate saves will always be unique.
            is_duplicate_split = (t_key and t_key in seen_entry_times) or (content_key in seen_contents)
            
            if t_key: seen_entry_times.add(t_key)
            seen_contents.add(content_key)

            # Metadata formatting (re-mapped for display later)
            entry_time_display = self._safe_format_datetime(entry.get('EntryTime', ''), '%Y-%m-%d %H:%M:%S')

            # Worker extraction and WorkTime determination (for display)
            consolidated_workers = c_workers
            display_worktime = ""
            
            # Worker filtering logic
            if filter_worker != '전체':
                consolidated_workers = filter_worker
                f_w_c = filter_worker.replace(' ', '').lower()
                for j in range(1, 11):
                    u_k = 'User' if j == 1 else f'User{j}'
                    u_v_raw = clean_s(entry.get(u_k, ''))
                    u_v_c = marker_pattern.sub('', u_v_raw.replace(' ', '').lower()).strip()
                    if u_v_c == f_w_c:
                        wt_k = 'WorkTime' if j == 1 else f'WorkTime{j}'
                        display_worktime = clean_s(entry.get(wt_k, ''))
                        break
            else:
                display_worktime = c_worktime

            # Numeric calculations
            def to_f_local(v):
                try:
                    if pd.isna(v) or str(v).lower() == 'nan': return 0.0
                    return float(str(v).replace(',', ''))
                except: return 0.0

            q_val = to_f_local(entry.get('검사량', entry.get('수량', 0.0)))
            p_val = to_f_local(entry.get('단가', 0.0))
            t_val_cost = to_f_local(entry.get('출장비', 0.0))
            m_val_cost = to_f_local(entry.get('일식', 0.0))
            f_val_cost = to_f_local(entry.get('검사비', 0.0))
            
            # Sum totals (Guarded by duplicate check to prevent double-counting)
            if not is_duplicate_split:
                total_test_amount += q_val
                total_unit_price += p_val
                total_travel_cost += t_val_cost
                total_meal_cost += m_val_cost
                total_test_fee += f_val_cost
                total_film_count += to_f_local(entry.get('FilmCount', 0.0))

            # Cumulative mileage (Always sum)
            milk = to_f_local(entry.get('주행거리', entry.get('거리', 0)))
            total_mileage += milk
            if milk > 0.001:
                min_mileage = min(min_mileage, milk) if min_mileage != float('inf') else milk
                max_mileage = max(max_mileage, milk) if max_mileage != float('-inf') else milk

            # OT hours and amounts calculation
            row_ot_hours = 0.0
            row_ot_amount = 0
            row_ots = []
            
            # Exact column name matching for each worker slot
            site_pairs = []
            all_keys = set(entry.keys())
            for j in range(1, 11):
                # j=1: 'User', j=2: 'User2', etc.
                uk = ('User' if j == 1 else f'User{j}') if ('User' if j == 1 else f'User{j}') in all_keys else None
                wk = ('WorkTime' if j == 1 else f'WorkTime{j}') if ('WorkTime' if j == 1 else f'WorkTime{j}') in all_keys else None
                ok = ('OT' if j == 1 else f'OT{j}') if ('OT' if j == 1 else f'OT{j}') in all_keys else None
                if uk:
                    site_pairs.append((uk, wk, ok))
            
            for i in range(1, 11):
                if i <= len(site_pairs):
                    uk, wk, ok = site_pairs[i-1]
                    uv = clean_s(entry.get(uk, ''))
                    
                    if not uv or (filter_worker != '전체' and uv != filter_worker):
                        row_ots.append("")
                        continue

                    ots = str(entry.get(ok, '')).strip()
                    wts = str(entry.get(wk, '')).strip()
                    
                    if ots and ots not in ('nan', '0.0', '0'):
                        try:
                            if '(' in ots and '원)' in ots:
                                h_p = float(ots.split('시간')[0])
                                a_p = int(_re.sub(r'[^0-9]', '', ots.split('(')[1].split('원')[0]))
                            elif ots.replace(',', '').isdigit():
                                a_p = int(ots.replace(',', ''))
                                h_p, _ = self._calculate_ot_from_worktime(wts, pd.to_datetime(entry.get('Date', datetime.datetime.now())))
                            else:
                                a_p = 0
                                h_p = self._parse_ot_hours(ots)
                                
                            # Activity-based hours: Max of all workers in this row
                            row_ot_hours = max(row_ot_hours, h_p)
                            # Cost-based amounts: Always sum across all workers
                            row_ot_amount += a_p
                            
                            if not is_duplicate_split:
                                # Update global individual totals (primarily for column data presence/amounts)
                                total_ot_amount += a_p
                                total_indiv_ot_hours[i-1] += h_p
                                total_indiv_ot_amounts[i-1] += a_p
                                
                            row_ots.append(f"{a_p:,}")
                        except:
                            row_ots.append(ots)
                    else:
                        row_ots.append("")  # Empty when no OT data (not '0')
                else: row_ots.append("")
            
            # Trim trailing empty OT slots so inactive workers don't create visible columns
            while row_ots and row_ots[-1] == "":
                row_ots.pop()
            # Re-pad to 10 with empty strings (the Treeview expects 10 slots)
            while len(row_ots) < 10:
                row_ots.append("")

            # Global OT Hours: Sum of per-activity maximums
            if not is_duplicate_split:
                total_ot_hours += row_ot_hours

            # Global Work Hours (ONLY for primary rows)
            if not is_duplicate_split and display_worktime and '~' in str(display_worktime):
                try:
                    cwt = marker_pattern.sub('', str(display_worktime)).strip()
                    if '~' in cwt:
                        pts = cwt.split('~')
                        sh, sm = map(int, pts[0].split(':'))
                        eh, em = map(int, pts[1].split(':'))
                        sm_t = sh * 60 + sm
                        em_t = eh * 60 + em
                        if em_t < sm_t: em_t += 1440
                        total_work_hours += (em_t - sm_t) / 60.0
                except: pass

            # RTK values
            def robust_to_f(v):
                if pd.isna(v) or str(v).lower() in ('nan', 'none', ''): return 0.0
                try:
                    cl = _re.sub(r'[^0-9\.\-]', '', str(v))
                    return float(cl) if cl else 0.0
                except: return 0.0

            rtk_vals = []
            row_rtk_sum = 0.0
            for i, cat in enumerate(['센터미스', '농도', '마킹미스', '필름마크', '취급부주의', '고객불만', '기타']):
                v = robust_to_f(entry.get(f'RTK_{cat}', 0))
                rtk_vals.append(f"{v:.1f}" if abs(v) > 0.001 else "")
                row_rtk_sum += v
                if not is_duplicate_split: total_rtk[i] += v
            rtk_vals.append(f"{row_rtk_sum:.1f}" if abs(row_rtk_sum) > 0.001 else "")
            if not is_duplicate_split: total_rtk[7] += row_rtk_sum
            
            # NDT values
            ndt_vals = []
            for i, mat in enumerate(["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]):
                v = robust_to_f(entry.get(f'NDT_{mat}'))
                if v == 0:
                    if mat == "흑색자분": v = robust_to_f(entry.get('NDT_자분'))
                    elif mat == "백색페인트": v = robust_to_f(entry.get('NDT_페인트'))
                    elif mat == "형광침투제": v = robust_to_f(entry.get('NDT_형광'))
                ndt_vals.append(f"{v:.1f}" if abs(v) > 0.001 else "")
                if not is_duplicate_split: total_ndt[i] += v

            # Remarks and visibility checks
            note_raw = str(entry.get('Note', '')).strip() if not pd.isna(entry.get('Note')) else ""
            display_remark = f"[{consolidated_workers}] {note_raw}" if consolidated_workers and note_raw else note_raw
            
            def has_content(v):
                if pd.isna(v) or str(v).lower() in ('nan', 'none', '') or str(v).strip() in ('', '-', '0', '0.0', '0시간'): return False
                return True
            
            if has_content(entry.get('장비명')): has_equip = True
            if has_content(entry.get('검사방법')): has_method = True
            if has_content(note_raw): has_note = True
            if has_content(entry.get('MaterialID')): has_product = True
            if has_content(entry.get('EntryTime')): has_entry_time = True
            if has_content(entry.get('차량번호')): has_vehicle_no = True
            if has_content(entry.get('주행거리')) or has_content(entry.get('거리')): has_mileage = True
            if has_content(entry.get('차량점검')): has_veh_insp = True
            if has_content(entry.get('차량비고')): has_veh_note = True
            if has_content(entry.get('회사코드')): has_co_code = True
            if has_content(entry.get('업체명')): has_company = True
            if has_content(entry.get('적용코드')): has_applied_code = True
            if has_content(entry.get('검사품명')): has_insp_item = True
            if has_content(entry.get('성적서번호')): has_report_no = True
            if has_content(entry.get('일식')): has_meal_cost = True
            if has_content(entry.get('Unit')): has_unit = True
            
            # [VISUAL DEDUPING] If duplicate split row, clear numeric values for cleaner view/export
            disp_q = f"{q_val:.1f}"
            disp_p = f"{p_val:,.0f}"
            disp_t = f"{t_val_cost:,.0f}"
            disp_f = f"{f_val_cost:,.0f}"
            disp_m = f"{m_val_cost:,.0f}"
            disp_oth = f"{row_ot_hours:.1f}"
            disp_ota = f"{row_ot_amount:,}"

            disp_row_ots = row_ots
            disp_rtk = rtk_vals
            disp_ndt = ndt_vals
            
            if is_duplicate_split:
                disp_q = ""
                disp_p = ""
                disp_t = ""
                disp_m = ""
                disp_f = ""
                disp_oth = ""
                disp_ota = ""
                disp_row_ots = [""] * 10
                disp_rtk = [""] * 8
                disp_ndt = [""] * 7
            
            v_tuple = (
                usage_date,
                entry.get('업체명', ''),
                entry.get('적용코드', ''),
                entry.get('Site', ''),
                entry.get('검사품명', ''),
                entry.get('성적서번호', ''),
                consolidated_workers, # Header: '작업자' (Index 6)
                display_worktime,     # Header: '작업시간' (Index 7)
                *disp_row_ots,        # OT1..OT10 (Index 8..17)
                entry.get('장비명', ''), # Header: '장비명' (Index 18)
                entry.get('검사방법', ''),# Header: '검사방법' (Index 19)
                entry.get('회사코드', ''), # Index 20
                disp_q,               # Index 21 (수량)
                entry.get('Unit', ''), # Index 22 (단위) [NEW]
                disp_p,               # Index 23 (단가)
                disp_t,               # Index 23 (출장비)
                disp_m,               # Index 24 (일식)
                disp_f,               # Index 25 (검사비)
                disp_oth,             # Index 26 (OT시간)
                disp_ota,             # Index 27 (OT금액)
                mat_name,             # Index 28 (품목명)
                *disp_rtk,            # Index 29..36
                *disp_ndt,            # Index 37..43
                display_remark,       # Index 44
                str(entry.get('EntryTime', '')),
                self.clean_nan(entry.get('차량번호', '')),
                self.clean_nan(entry.get('주행거리', '')),
                self.clean_nan(entry.get('차량점검', '')),
                self.clean_nan(entry.get('차량비고', '')),
                ", ".join(raw_workers)
            )


            while len(v_tuple) < 48: v_tuple += ("",)
            self.daily_usage_tree.insert('', tk.END, values=v_tuple, tags=(str(idx),))
            
        # Insert last daily subtotal and final total row if data exists
        if not filtered_df.empty:
            
            # Final overall total
            self.daily_usage_tree.tag_configure('total', background='#E8F4F8', font=('Arial', 9, 'bold'))
            
            total_rtk_ready = [f"{v:.1f}" if abs(v) > 0.001 else "" for v in total_rtk]
            while len(total_rtk_ready) < 8: total_rtk_ready.append("")
            
            total_ndt_ready = [f"{v:.1f}" if abs(v) > 0.001 else "" for v in total_ndt]
            while len(total_ndt_ready) < 7: total_ndt_ready.append("")

            total_values = [
                '--- 전체 누계 ---',
                '', # 업체명 (1)
                '', # 적용코드 (2)
                '', # 현장 (3)
                '', # 검사품명 (4)
                '', # 성적서번호 (5)
                '', # 작업자 (6)
                f"{total_work_hours:.1f} Hrs" if total_work_hours > 0.001 else "", # 작업시간 (7)
                # Individual OT Totals (Simplified: Amount only)
                *[f"{a:,}" if a > 0.001 else "" for a in total_indiv_ot_amounts],
                '', # 장비명
                '', # 검사방법
                '', # 회사코드
                f"{total_test_amount:.1f}" if total_test_amount > 0.001 else "",
                '', # 단위 (Unit)
                '', # 단가 (Unit Price is not summed)
                f"{total_travel_cost:,.0f}" if total_travel_cost > 0.001 else "",
                f"{total_meal_cost:,.0f}" if total_meal_cost > 0.001 else "", # Added back missing index
                f"{total_test_fee:,.0f}" if total_test_fee > 0.001 else "",
                f"{total_ot_hours:.1f}" if total_ot_hours > 0.001 else "", # OT시간 합계
                f"{total_ot_amount:,}" if total_ot_amount > 0.001 else "",  # OT금액 합계
                '', # 품목명
                *total_rtk_ready,
                *total_ndt_ready,
                '',   # 비고
                '',   # 입력시간
                '',   # 차량번호
                f"누계: {total_mileage:,.1f} km" if total_mileage > 0.001 else "0 km",   # 주행거리 합계
                '',   # 차량점검
                '',   # 차량비고
                ''    # (Full작업자)
            ]
            # [DEFENSIVE] Ensure Total row matches header count exactly
            while len(total_values) < 48: total_values.append("")
            self.daily_usage_tree.tag_configure('total', background='#E8F4F8', font=('Arial', 12, 'bold'))
            self.daily_usage_tree.insert('', tk.END, values=total_values, tags=('total',))
            
            # --- Dynamic Column Hiding ---
            # Mandatory cols (always show to maintain core row identity)
            # [USER REQUEST] Force '센터미스' and '농도' to be mandatory/always shown
            mandatory_cols = ['날짜', '현장', '작업자']
            
            # Use a slightly more robust threshold for data presence
            # Also handle potential string '0.0' leftovers and common empty markers
            def is_active(val):
                if val is None: return False
                s = str(val).strip().lower()
                # Explicitly list all known 'zero' or 'empty' string representations
                if s in ('', '0', '0.0', '0.00', 'nan', 'none', '-', '0.0시간', '0.0(0원)', '0.0 (0원)', '0시간', '0원', '0(0원)', '0 (0원)'):
                    return False
                try:
                    # Robust cleaning: remove everything except numbers, dots, and minus
                    clean_s = re.sub(r'[^0-9\.\-]', '', s)
                    if not clean_s: return False # Only markers, but no number? Treat as empty.
                    v = float(clean_s)
                    return abs(v) > 0.001 # Slightly wider epsilon
                except:
                    # If it's a non-numeric string (like a note), return True if not empty
                    return bool(s)
            
            dynamic_col_status = {
                '작업시간': is_active(total_work_hours),
                '수량': is_active(total_test_amount),
                '단가': is_active(total_unit_price),
                '출장비': is_active(total_travel_cost),
                '검사비': is_active(total_test_fee),
                'OT시간': is_active(total_ot_hours),
                'OT금액': is_active(total_ot_amount),
                '장비명': has_equip,
                '검사방법': has_method,
                '비고': has_note,
                '차량번호': has_vehicle_no,
                '주행거리': has_mileage,
                '차량점검': has_veh_insp,
                '차량비고': has_veh_note,
                '품목명': has_product, # Now content-based
                '입력시간': has_entry_time, # Now content-based
                '회사코드': has_co_code,
                '업체명': has_company,
                '적용코드': has_applied_code,
                '검사품명': has_insp_item,
                '성적서번호': has_report_no,
                '일식': has_meal_cost,
                '단위': has_unit,
                '수량': is_active(total_test_amount), # [FIX] Ensure these are explicitly handled
                '단가': is_active(total_unit_price),
                '출장비': is_active(total_travel_cost),
                '검사비': is_active(total_test_fee)
            }
            
            # Individual OT columns
            for i in range(1, 11):
                col_name = f'OT{i}'
                dynamic_col_status[col_name] = is_active(total_indiv_ot_amounts[i-1])
            
            # RTK Columns
            rtk_col_names = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타"]
            for i, col_name in enumerate(rtk_col_names):
                active_status = is_active(total_rtk[i])
                # [USER REQUEST] Force hide '마킹미스' and others if they have no significant data
                if col_name in ["마킹미스", "필름마크", "취급부주의", "고객불만", "기타"]:
                    dynamic_col_status[col_name] = active_status
                else:
                    # '센터미스' and '농도' should follow mandatory_cols but we safeguard here too
                    dynamic_col_status[col_name] = active_status
            
            # Diagnostic Log (visible in terminal for aid)
            # print(f"DEBUG RTK Totals: {dict(zip(rtk_col_names, total_rtk[:7]))}, Total: {total_rtk[7]}")
            
            dynamic_col_status['RTK총계'] = is_active(total_rtk[7])
            
            # NDT Columns
            ndt_col_names = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
            for i, col_name in enumerate(ndt_col_names):
                dynamic_col_status[col_name] = is_active(total_ndt[i])
                
            # --- BUILD FINAL VISIBILITY (MERGED LOGIC) ---
            all_cols = list(self.daily_usage_tree['columns'])
            # manual_set: columns user explicitly hid via column manager (empty = no exclusions)
            manual_hidden = set(all_cols) - set(getattr(self, 'manual_visible_cols', all_cols))
            
            # [SAFETY] Core columns that should almost never be hidden unless user is very specific
            # [REFINED] Minimal mandatory columns to allow smarter auto-hiding of empty fields
            mandatory_cols = ['날짜', '현장', '작업자']
            
            final_visible = []
            # We iterate in ALL_COLS order to maintain the original column sequence
            for col in all_cols:
                if col == '(Full작업자)': continue
                
                # 1. Mandatory override: Always show core columns
                if col in mandatory_cols:
                    final_visible.append(col)
                    continue
                
                # 2. SMART HIDING: For columns we track data status for
                if col in dynamic_col_status:
                    # Show ONLY if it has data AND isn't manually hidden
                    if dynamic_col_status[col] and col not in manual_hidden:
                        final_visible.append(col)
                    # Skip untracked fallback for this column (i.e. if NO data, it's HIDDEN)
                    continue
                
                # 3. Fallback for untracked columns (e.g. manually added custom columns)
                if col not in manual_hidden:
                    final_visible.append(col)

            # [STABILITY] Clear the Treeview's displayed columns
            self.daily_usage_tree['displaycolumns'] = final_visible
            

            
            # Header renames (if any)

            # Ensure stretch=False and minwidth is relaxed for ALL displayed columns
            # This is critical for making all columns, especially the last one, resizable
            for col in final_visible:
                self.daily_usage_tree.column(col, stretch=False, minwidth=20)

            # Re-apply saved column widths if available
            if hasattr(self, 'tab_config') and 'daily_usage_col_widths' in self.tab_config:
                saved_widths = self.tab_config['daily_usage_col_widths']
                for col in final_visible:
                    if col in saved_widths:
                        try:
                            self.daily_usage_tree.column(col, width=int(saved_widths[col]), stretch=False)
                            # Enforce minimums for high-precision cols to prevent truncation
                            if col == '날짜':
                                self.daily_usage_tree.column(col, width=max(int(saved_widths[col]), 160), stretch=False)
                            elif col == '입력시간':
                                self.daily_usage_tree.column(col, width=max(int(saved_widths[col]), 300), stretch=False)
                        except: pass
            
            # Ensure Total Row stays at bottom
            self.daily_usage_tree.detach(self.daily_usage_tree.get_children()[-1])
            self.daily_usage_tree.insert('', tk.END, values=total_values, tags=('total',))
        else:
            # If empty, show only mandatory columns to keep the view clean
            all_cols = self.daily_usage_tree['columns']
            final_visible = ['날짜', '현장', '작업자']
            self.daily_usage_tree['displaycolumns'] = final_visible
            
            # Standard column setup for display
            for col in final_visible:
                self.daily_usage_tree.column(col, stretch=False, minwidth=20)
                
            # Re-apply saved column widths if available
            if hasattr(self, 'tab_config') and 'daily_usage_col_widths' in self.tab_config:
                saved_widths = self.tab_config['daily_usage_col_widths']
                for col in final_visible:
                    if col in saved_widths:
                        try:
                            self.daily_usage_tree.column(col, width=int(saved_widths[col]), stretch=False)
                        except: pass

    def reset_daily_usage_filters(self):
        """Reset all daily usage history filters to default values"""
        try:
            # 1. Reset Dates
            start_date = datetime.datetime(2024, 1, 1)
            self.ent_daily_start_date.set_date(start_date)
            self.ent_daily_end_date.set_date(datetime.datetime.now())
            
            # 2. Reset Comboboxes
            filter_combos = [
                self.cb_daily_filter_company, self.cb_daily_filter_site, 
                self.cb_daily_filter_material, self.cb_daily_filter_equipment,
                self.cb_daily_filter_worker, self.cb_daily_filter_vehicle,
                self.cb_daily_filter_shift
            ]
            for combo in filter_combos:
                if hasattr(self, combo.winfo_name()) or True: # Safeguard
                    combo.set('전체')
            
            # 3. Refresh View
            self.update_daily_usage_view()
            
        except Exception as e:
            print(f"Error resetting daily usage filters: {e}")

    def _auto_adjust_tree_columns(self, tree, expand_only=False):
        """Automatically adjust column widths to fit content"""
        import tkinter.font as tkfont
        
        # Use the actual font used in the Treeview (12pt Malgun Gothic)
        # This ensures measurement matches display perfectly.
        content_font = tkfont.Font(family="Malgun Gothic", size=12)
        heading_font = tkfont.Font(family="Malgun Gothic", size=12, weight="bold")
        
        # We only care about visible columns
        visible_cols = tree['displaycolumns']
        if not visible_cols or visible_cols == ('#all'):
             visible_cols = tree['columns']

        for col in visible_cols:
            # Maximum Overkill: Character Count Heuristic
            # Assume 40px per character + 200px buffer
            # This GUARANTEES visibility in any environment
            
            # Measure heading
            w = len(col) * 40 + 200
            
            # Measure content
            col_index = list(tree['columns']).index(col)
            for item in tree.get_children():
                val = tree.item(item, 'values')
                if val and col_index < len(val):
                    text_len = len(str(val[col_index]))
                    content_w = text_len * 40 + 200
                    if content_w > w:
                        w = content_w
            
            # Apply specific constraints
            # Ultra-Safe: No max limit for potentially long text columns
            if col in ['품목명', '작업자', '비고']:
                w = max(w, 300) # Minimum base
                if col == '품목명': w = max(w, 700)
                elif col == '작업자': w = max(w, 300)
                elif col == '비고': w = max(w, 600)
            else:
                # Other columns get a relaxed global max
                min_w = 200
                if col == '날짜': min_w = 400
                elif col == '현장': min_w = 300
                elif col == '장비명': min_w = 250
                elif col == '검사방법': min_w = 200
                
                w = max(min_w, min(w, 2000)) 
            
            # Smart Auto-Expand: Only grow if content requires it
            if expand_only:
                current_w = int(tree.column(col, 'width'))
                if w <= current_w:
                    continue  # Keep user's manual adjustment
            
            # All columns are now user-resizable
            tree.column(col, width=w, minwidth=20, stretch=False, anchor='center')
        
        # Force layout update
        tree.update_idletasks()

    def delete_daily_usage_entry(self):
        """선택된 일일 사용량 기록 삭제 (인덱스 기반으로 정확하게 삭제 및 재고 환원)"""
        selected_items = self.daily_usage_tree.selection()
        
        if not selected_items:
            messagebox.showwarning("선택 오류", "삭제할 항목을 선택해주세요.")
            return
        
        # 사용자 확인
        result = messagebox.askyesno("삭제 확인", f"{len(selected_items)}개의 기록을 삭제하시겠습니까?\n(삭제 시 차감되었던 재고도 자동으로 환원됩니다.)")
        
        if not result:
            return
        
        indices_to_delete = []
        for item in selected_items:
            tags = self.daily_usage_tree.item(item, 'tags')
            if tags:
                try:
                    df_idx = int(tags[0])
                    indices_to_delete.append(df_idx)
                except ValueError:
                    continue
        
        if not indices_to_delete:
            messagebox.showwarning("삭제 실패", "삭제할 항목의 데이터 정보를 찾을 수 없습니다.")
            return

        deleted_count = 0
        try:
            # Sort indices in descending order to avoid index shift issues during deletion
            # But wait, we are using boolean masking or actual indices in the original dataframe
            # Since we will drop them at once or one by one from the original, it's better to collect all indices.
            
            for idx in indices_to_delete:
                if idx not in self.daily_usage_df.index:
                    continue
                
                entry = self.daily_usage_df.loc[idx]
                
                # --- [FIX] 재고 환원 로직 (자동 차감된 트랜잭션만 선택적으로 환원) ---
                site = entry.get('Site', '')
                usage_date = pd.to_datetime(entry.get('Date'))
                
                if not self.transactions_df.empty:
                    # 해당 현장/날짜의 "(자동 차감)" 트랜잭션을 찾아 삭제
                    trans_mask = (
                        (pd.to_datetime(self.transactions_df['Date'], errors='coerce').dt.normalize() == pd.to_datetime(usage_date).normalize()) &
                        (self.transactions_df['Site'].astype(str) == str(site)) &
                        (self.transactions_df['Type'] == 'OUT') &
                        (self.transactions_df['Note'].str.contains(f"{site} 현장 사용", na=False))
                    )
                    self.transactions_df = self.transactions_df[~trans_mask]
                    deleted_count += 1

            # 기록 삭제
            self.daily_usage_df = self.daily_usage_df.drop(indices_to_delete)
            
            # 인덱스 초기화는 하지 않음 (태그 매칭을 위해 원본 인덱스 유지 권장하나, 
            # drop 후에는 뷰 업데이트 시 어차피 다시 생성되므로 안전하게 reset_index 가능)
            self.daily_usage_df = self.daily_usage_df.reset_index(drop=True)
            
            if self.save_data():
                messagebox.showinfo("삭제 완료", f"{len(indices_to_delete)}개의 기록이 삭제되고 재고가 환원되었습니다.")
                self.update_daily_usage_view()
                self.update_stock_view()
                self.update_transaction_view()
                self.refresh_inquiry_filters()
        except Exception as e:
            messagebox.showerror("삭제 오류", f"기록 삭제 중 오류가 발생했습니다: {e}")

    def open_edit_daily_usage_dialog(self):
        """Open a dialog to edit the selected daily usage record"""
        selection = self.daily_usage_tree.selection()
        if not selection:
            messagebox.showwarning("선택 오류", "수정할 항목을 선택해주세요.")
            return
            
        item = self.daily_usage_tree.item(selection[0])
        tags = item.get('tags', [])
        if not tags:
            messagebox.showwarning("데이터 오류", "선택한 항목의 데이터 정보를 찾을 수 없습니다.")
            return
            
        try:
            df_idx = int(tags[0])
            entry_data = self.daily_usage_df.loc[df_idx]
        except (ValueError, KeyError, IndexError):
            messagebox.showwarning("데이터 오류", "데이터를 불러오는 중 오류가 발생했습니다.")
            return

        # Create Edit Dialog
        edit_win = tk.Toplevel(self.root)
        edit_win.title("일일 사용 기록 수정")
        edit_win.geometry("800x900")
        edit_win.transient(self.root)
        edit_win.grab_set()

        # Main scrollable frame
        container = ttk.Frame(edit_win)
        container.pack(expand=True, fill='both')
        
        canvas = tk.Canvas(container)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas, padding=20)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Scroll binding is handled globally in MaterialManager.__init__

        fields = {}
        
        # 1. Basic Info
        basic_frame = ttk.LabelFrame(scrollable_frame, text="기본 정보", padding=10)
        basic_frame.pack(fill='x', pady=5)
        
        # Date
        ttk.Label(basic_frame, text="날짜:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        ent_date = DateEntry(basic_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        curr_date = entry_data.get('Date', datetime.datetime.now().strftime('%Y-%m-%d'))
        try: ent_date.set_date(pd.to_datetime(curr_date))
        except: pass
        ent_date.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        fields['Date'] = ent_date

        # Site
        ttk.Label(basic_frame, text="현장:").grid(row=0, column=2, padx=5, pady=5, sticky='w')
        cb_site = ttk.Combobox(basic_frame, width=20, values=self.sites)
        cb_site.set(self.clean_nan(entry_data.get('Site', '')))
        cb_site.grid(row=0, column=3, padx=5, pady=5, sticky='w')
        fields['Site'] = cb_site
        
        # Company & Applied Code
        ttk.Label(basic_frame, text="업체명:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
        cb_company = ttk.Combobox(basic_frame, width=20, values=self.companies)
        cb_company.set(self.clean_nan(entry_data.get('업체명', '')))
        cb_company.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        fields['업체명'] = cb_company
        
        ttk.Label(basic_frame, text="적용코드:").grid(row=1, column=2, padx=5, pady=5, sticky='w')
        ent_code = ttk.Entry(basic_frame, width=12)
        ent_code.insert(0, self.clean_nan(entry_data.get('적용코드', '')))
        ent_code.grid(row=1, column=3, padx=5, pady=5, sticky='w')
        fields['적용코드'] = ent_code

        # Material
        ttk.Label(basic_frame, text="품목명:").grid(row=2, column=0, padx=5, pady=5, sticky='w')
        cb_mat = ttk.Combobox(basic_frame, width=40)
        mat_options = []
        for _, m_row in self.materials_df.iterrows():
            if m_row.get('Active', 1) == 0: continue
            mat_options.append(self.get_material_display_name(m_row['MaterialID']))
        
        # Deduplicate and sort for a clean list
        mat_options = sorted(list(set([m for m in mat_options if m])))
        cb_mat['values'] = mat_options
        
        # Determine current display name
        curr_mat_id = entry_data.get('MaterialID')
        curr_mat_display = self.get_material_display_name(curr_mat_id)
        
        cb_mat.set(curr_mat_display)
        cb_mat.grid(row=2, column=1, columnspan=3, padx=5, pady=5, sticky='w')
        fields['Material'] = cb_mat

        # Equipment & Method
        ttk.Label(basic_frame, text="장비명:").grid(row=3, column=0, padx=5, pady=5, sticky='w')
        cb_equip = ttk.Combobox(basic_frame, width=20, values=self.equipments)
        cb_equip.set(self.clean_nan(entry_data.get('장비명', '')))
        cb_equip.grid(row=3, column=1, padx=5, pady=5, sticky='w')
        fields['장비명'] = cb_equip

        ttk.Label(basic_frame, text="검사방법:").grid(row=3, column=2, padx=5, pady=5, sticky='w')
        cb_method = ttk.Combobox(basic_frame, width=10, values=['RT', 'PAUT', 'UT', 'MT', 'PT', 'ETC'])
        cb_method.set(self.clean_nan(entry_data.get('검사방법', '')))
        cb_method.grid(row=3, column=3, padx=5, pady=5, sticky='w')
        fields['검사방법'] = cb_method
        
        # [V13_FIX] Add Unit to basic info - Now using dynamic list from unit management
        ttk.Label(basic_frame, text="단위:").grid(row=3, column=4, padx=5, pady=5, sticky='w')
        cb_unit = ttk.Combobox(basic_frame, width=10, values=self.daily_units)
        cb_unit.set(self.clean_nan(entry_data.get('Unit', entry_data.get('단위', '매'))))
        cb_unit.grid(row=3, column=5, padx=5, pady=5, sticky='w')
        fields['Unit'] = cb_unit
        
        # Inspection Item
        ttk.Label(basic_frame, text="검사품명:").grid(row=4, column=0, padx=5, pady=5, sticky='w')
        ent_item = ttk.Entry(basic_frame, width=40)
        ent_item.insert(0, self.clean_nan(entry_data.get('검사품명', '')))
        ent_item.grid(row=4, column=1, columnspan=3, padx=5, pady=5, sticky='w')
        fields['검사품명'] = ent_item

        # 2. Quantities & Costs
        cost_frame = ttk.LabelFrame(scrollable_frame, text="수량 및 비용", padding=10)
        cost_frame.pack(fill='x', pady=5)
        
        qty_configs = [
            ('검사량', '검사량'), ('단가', '단가'), ('출장비', '출장비'),
            ('일식', '일식'), ('검사비', '검사비'), ('필름매수', 'FilmCount')
        ]
        for i, (lbl, key) in enumerate(qty_configs):
            row = i // 3
            col = (i % 3) * 2
            ttk.Label(cost_frame, text=f"{lbl}:").grid(row=row, column=col, padx=5, pady=5, sticky='w')
            ent = ttk.Entry(cost_frame, width=12)
            val = entry_data.get(key, 0)
            if pd.isna(val): val = 0
            ent.insert(0, str(val))
            ent.grid(row=row, column=col+1, padx=5, pady=5, sticky='w')
            fields[key] = ent

        # 3. RTK & NDT
        material_frame = ttk.Frame(scrollable_frame)
        material_frame.pack(fill='x', pady=5)
        
        rtk_frame = ttk.LabelFrame(material_frame, text="RTK 사용량 (재촬영)", padding=10)
        rtk_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))
        
        rtk_cats = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타"]
        for i, cat in enumerate(rtk_cats):
            ttk.Label(rtk_frame, text=f"{cat}:").grid(row=i, column=0, padx=5, pady=2, sticky='w')
            ent = ttk.Entry(rtk_frame, width=10)
            val = entry_data.get(f'RTK_{cat}', 0)
            if pd.isna(val): val = 0
            ent.insert(0, str(val))
            ent.grid(row=i, column=1, padx=5, pady=2, sticky='w')
            fields[f'RTK_{cat}'] = ent

        ndt_frame = ttk.LabelFrame(material_frame, text="NDT 약품 사용량", padding=10)
        ndt_frame.pack(side='left', fill='both', expand=True, padx=(5, 0))
        
        ndt_cats = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
        for i, cat in enumerate(ndt_cats):
            ttk.Label(ndt_frame, text=f"{cat}:").grid(row=i, column=0, padx=5, pady=2, sticky='w')
            ent = ttk.Entry(ndt_frame, width=10)
            m_key = "".join(cat.split())
            val = entry_data.get(f'NDT_{m_key}', 0)
            if pd.isna(val): val = 0
            ent.insert(0, str(val))
            ent.grid(row=i, column=1, padx=5, pady=2, sticky='w')
            fields[f'NDT_{m_key}'] = ent

            fields[f'NDT_{m_key}'] = ent
            
        # 4. Vehicle Info
        vehicle_frame = ttk.LabelFrame(scrollable_frame, text="차량 점검 정보", padding=10)
        vehicle_frame.pack(fill='x', pady=5)
        
        v_configs = [
            ('차량번호', '차량번호'), ('주행거리', '주행거리'),
            ('점검내용', '차량점검'), ('차량비고', '차량비고')
        ]
        for v_i, (v_lbl, v_key) in enumerate(v_configs):
            v_row = v_i // 2
            v_col = (v_i % 2) * 2
            ttk.Label(vehicle_frame, text=f"{v_lbl}:").grid(row=v_row, column=v_col, padx=5, pady=5, sticky='w')
            v_ent = ttk.Entry(vehicle_frame, width=25)
            v_val = entry_data.get(v_key, '')
            v_ent.insert(0, self.clean_nan(v_val))
            v_ent.grid(row=v_row, column=v_col+1, padx=5, pady=5, sticky='w')
            fields[v_key] = v_ent

        # 5. Workers
        worker_main_frame = ttk.LabelFrame(scrollable_frame, text="작업자 정보 (1~10)", padding=10)
        worker_main_frame.pack(fill='x', pady=5)
        


        # [NEW] Inner-dialog helper for Fee calculation
        def update_edit_fee_calc(event=None):
            try:
                def get_f(val):
                    try:
                        v = str(val).strip().replace(',', '')
                        return float(v) if v else 0.0
                    except: return 0.0

                qty = get_f(fields['검사량'].get())
                price = get_f(fields['단가'].get())
                travel = get_f(fields['출장비'].get())
                meal = get_f(fields['일식'].get())
                
                calc_fee = (qty * price) + travel + meal
                fields['검사비'].delete(0, tk.END)
                fields['검사비'].insert(0, f"{calc_fee:,.0f}")
            except: pass

        # Bindings for auto-calc
        fields['검사량'].bind('<KeyRelease>', lambda e: update_edit_fee_calc())
        
        # [NEW] Add comma auto-formatting for edit fields
        for k in ['단가', '출장비', '일식', '검사비']:
            if k in fields:
                fields[k].bind('<KeyRelease>', update_edit_fee_calc, add='+')
                fields[k].bind('<FocusOut>', lambda e, widget=fields[k]: self.format_entry_with_commas(e, widget), add='+')
                fields[k].bind('<Return>', lambda e, widget=fields[k]: self.format_entry_with_commas(e, widget), add='+')
        
        # Ensure FilmCount field also formats nicely (integer)
        if 'FilmCount' in fields:
            fields['FilmCount'].bind('<FocusOut>', lambda e, widget=fields['FilmCount']: self.format_entry_with_commas(e, widget), add='+')

        # Grid for workers: 5 rows x 2 cols
        worker_fields = {}
        for i in range(1, 11):
            w_row = (i-1) // 2
            w_col = (i-1) % 2
            
            w_frame = ttk.Frame(worker_main_frame, padding=2)
            w_frame.grid(row=w_row, column=w_col, padx=10, pady=5, sticky='nw')
            
            ttk.Label(w_frame, text=f"{i}:", width=2).pack(side='left')
            
            # Name
            cb_name = ttk.Combobox(w_frame, width=8, values=self.users)
            name_key = 'User' if i == 1 else f'User{i}'
            cb_name.set(self.clean_nan(entry_data.get(name_key, '')))
            cb_name.pack(side='left', padx=2)
            worker_fields[f'name{i}'] = cb_name
            
            # Time
            cb_time = ttk.Combobox(w_frame, width=16, values=self.worktimes)
            time_key = 'WorkTime' if i == 1 else f'WorkTime{i}'
            cb_time.set(self.clean_nan(entry_data.get(time_key, '')))
            cb_time.pack(side='left', padx=2)
            worker_fields[f'time{i}'] = cb_time
            
            # OT
            ent_ot = ttk.Entry(w_frame, width=18)
            ot_key = 'OT' if i == 1 else f'OT{i}'
            ent_ot.insert(0, self.clean_nan(entry_data.get(ot_key, '')))
            ent_ot.pack(side='left', padx=2)
            worker_fields[f'ot{i}'] = ent_ot
            
            # Auto-calculate OT for each worker when time changes
            def on_time_changed(event, idx=i, w_ent=cb_time, ot_ent=ent_ot, d_ent=ent_date):
                self.calculate_and_update_ot_manual(w_ent.get(), ot_ent, d_ent.get_date())
            
            cb_time.bind('<Return>', on_time_changed)
        
        # [NEW] Apply autocomplete to main edit fields
        self.enable_autocomplete(cb_site, values_list_attr='sites')
        self.enable_autocomplete(cb_mat, values_list=mat_options)
        self.enable_autocomplete(cb_equip, values_list_attr='equipment_suggestions')

        # 5. Buttons
        btn_frame = ttk.Frame(scrollable_frame, padding=20)
        btn_frame.pack(fill='x')
        
        def save_edits():
            # Collect data
            new_data = {}
            new_data['Date'] = ent_date.get_date().strftime('%Y-%m-%d')
            new_data['Site'] = cb_site.get().strip()
            new_data['업체명'] = cb_company.get().strip()
            new_data['적용코드'] = ent_code.get().strip()
            new_data['검사품명'] = ent_item.get().strip()
            
            full_mat_name = cb_mat.get().strip()
            # [FIX] Robust MaterialID mapping: Try to find matching ID using standard display name generator
            found_id = None
            if full_mat_name:
                for _, m_row in self.materials_df.iterrows():
                    m_id_check = m_row['MaterialID']
                    if self.get_material_display_name(m_id_check) == full_mat_name:
                        found_id = m_id_check
                        break
                
                # [FIX] Only auto-register if it's a CONSUMABLE. Equipment should NOT be registered.
                if not found_id:
                    if self._is_consumable_material(full_mat_name, cb_method.get().strip()):
                        found_id = self.register_new_material(full_mat_name, warehouse='현장', 규격='수정등록')
                    else:
                        found_id = full_mat_name
            
            new_data['Material'] = full_mat_name
            new_data['MaterialID'] = found_id if found_id is not None else entry_data.get('MaterialID')
            
            new_data['장비명'] = cb_equip.get().strip()
            new_data['검사방법'] = cb_method.get().strip()
            new_data['Unit'] = cb_unit.get().strip() # [V13_FIX] Save Unit field
            
            # Numeric fields
            for key in ['검사량', '단가', '출장비', '일식', '검사비', 'FilmCount']:
                try: 
                    v_str = fields[key].get().strip().replace(',', '')
                    new_data[key] = float(v_str) if v_str else 0.0
                except: new_data[key] = 0.0
                
            for cat in rtk_cats:
                try: new_data[f'RTK_{cat}'] = float(fields[f'RTK_{cat}'].get())
                except: new_data[f'RTK_{cat}'] = 0.0
                
            for cat in ndt_cats:
                m_key = "".join(cat.split())
                try: new_data[f'NDT_{m_key}'] = float(fields[f'NDT_{m_key}'].get())
                except: new_data[f'NDT_{m_key}'] = 0.0
                
            for key in ['차량번호', '주행거리', '차량점검', '차량비고']:
                new_data[key] = fields[key].get().strip()
                
            for i in range(1, 11):
                name_key = 'User' if i == 1 else f'User{i}'
                time_key = 'WorkTime' if i == 1 else f'WorkTime{i}'
                ot_key = 'OT' if i == 1 else f'OT{i}'
                
                new_data[name_key] = worker_fields[f'name{i}'].get().strip()
                new_data[time_key] = worker_fields[f'time{i}'].get().strip()
                new_data[ot_key] = worker_fields[f'ot{i}'].get().strip()

            # Provide immediate feedback
            btn_save_edit.config(state='disabled', text="저장 중...")
            edit_win.update_idletasks()
            
            try:
                if self.save_daily_usage_edits(df_idx, new_data):
                    edit_win.destroy()
                    messagebox.showinfo("수정 완료", "기록이 성공적으로 수정되었으며 재고 내역이 업데이트되었습니다.")
                else:
                    btn_save_edit.config(state='normal', text="수정사항 저장")
            except Exception as e:
                btn_save_edit.config(state='normal', text="수정사항 저장")
                messagebox.showerror("오류", f"수정 중 오류가 발생했습니다: {e}")

        btn_save_edit = ttk.Button(btn_frame, text="수정사항 저장", style='Big.TButton', command=save_edits)
        btn_save_edit.pack(side='right', padx=10)
        ttk.Button(btn_frame, text="취소", command=edit_win.destroy).pack(side='right', padx=10)

    def _calculate_ot_from_worktime(self, worktime_value, calculation_date):
        """Standalone helper to calculate OT hours and amount from a worktime string"""
        try:
            if not worktime_value:
                return 0.0, 0
                
            import re
            marker_pattern = MARKER_PATTERN
            clean_val = marker_pattern.sub('', str(worktime_value)).strip()
            
            # Use regex to find separator (~ or -) more robustly
            sep_match = re.search(r'[:\d]\s*([~-])\s*[:\d]', clean_val)
            if not sep_match: return 0.0, 0
            sep = sep_match.group(1)
            
            start_time_str, end_time_str = clean_val.split(sep)
            sh, sm = map(int, start_time_str.split(':'))
            eh, em = map(int, end_time_str.split(':'))
            
            start_f = sh + sm / 60.0
            end_f = eh + em / 60.0
            if end_f < start_f: end_f += 24
            total_duration = end_f - start_f
            if total_duration <= 0: return 0.0, 0

            weekday = calculation_date.weekday()
            is_holiday = weekday >= 5
            is_friday = (weekday == 4)

            ot_hours = 0.0
            amount = 0
            if is_holiday:
                ot_hours = total_duration
                amount = ot_hours * 7500
            else:
                if end_f > 18:
                    ot_start = max(start_f, 18.0)
                    ot_hours = end_f - ot_start
                    evening_end = min(end_f, 22.0)
                    evening_hours = max(0, evening_end - ot_start)
                    night_start = max(ot_start, 22.0)
                    night_end = min(end_f, 24.0)
                    night_hours = max(0, night_end - night_start)
                    dawn_start = max(ot_start, 24.0)
                    dawn_hours = max(0, end_f - dawn_start)
                    dawn_rate = 7500 if is_friday else 5000
                    amount = (evening_hours * 4000) + (night_hours * 5000) + (dawn_hours * dawn_rate)
            
            return ot_hours, int(amount)
        except:
            return 0.0, 0

    def calculate_and_update_ot_manual(self, worktime_value, ot_entry, calculation_date=None):
        """Helper for personal calculation during editing without complex widget lookups"""
        try:
            current_date = calculation_date if calculation_date else self.ent_daily_date.get_date()
            ot_hours, amount = self._calculate_ot_from_worktime(worktime_value, current_date)
            
            if amount > 0:
                ot_entry.delete(0, tk.END)
                ot_entry.insert(0, f"{amount:,}")
            elif ot_hours == 0:
                # If no OT, clear it or set to 0
                ot_entry.delete(0, tk.END)
                ot_entry.insert(0, "0")
        except: pass

    def _safe_set_daily_df(self, idx, col, val, is_numeric=False):
        """Defensively set a value in daily_usage_df handling all possible dtype conflicts"""
        if col not in self.daily_usage_df.columns:
            self.daily_usage_df.at[idx, col] = val
            return

        col_dtype = self.daily_usage_df[col].dtype
        is_stringy = pd.api.types.is_string_dtype(col_dtype) or pd.api.types.is_object_dtype(col_dtype)
        
        v_to_set = val
        if is_stringy and not is_numeric:
            # Force to cleaned string for string columns
            v_to_set = str(val).strip() if pd.notna(val) else ""
            if v_to_set in ('0.0', '0', 'nan', 'None', 'NaT'): v_to_set = ""
        elif is_numeric:
            # Force to float for numeric columns
            try:
                if isinstance(val, str):
                    v_to_set = float(val.replace(',', '').strip()) if val.strip() else 0.0
                else:
                    v_to_set = float(val) if pd.notna(val) else 0.0
            except:
                v_to_set = 0.0

        try:
            self.daily_usage_df.at[idx, col] = v_to_set
        except:
            try:
                # Type conflict (e.g. float into strict string column or vice versa)
                # Cast the column to object to allow the mixed/new type
                self.daily_usage_df[col] = self.daily_usage_df[col].astype(object)
                self.daily_usage_df.at[idx, col] = v_to_set
            except:
                # Final fallback
                self.daily_usage_df.at[idx, col] = str(v_to_set)

    def _is_consumable_material(self, name, method):
        """
        Determines if a material should be automatically registered and tracked as stock.
        Consumables: RT Films, MT/PT drugs, chemicals.
        Equipment (exclude): Scanners, Crawlers, Sources, Meters, Yokes, etc.
        """
        if not name: return False
        n = str(name).strip().upper().replace(' ', '')
        m = str(method).strip().upper()

        # 1. MT/PT consumables (NDT drugs) - Enhanced keywords
        ndt_keywords = [x.upper().replace(' ', '') for x in self.ndt_materials_all]
        ndt_keywords += ['PT약품', 'MT약품', 'NDT약품', '침투액', '세척액', '현상액', '자분액']
        
        # Stricter equipment check for MT/PT
        equip_keywords = ['YOKE', '장비', 'EQUIP', 'METER', 'GAUGE', 'UVLAMP', '전등', '라이트', '자화', 'SCANNER', '스캐너', 'CRAWLER', '크롤러']
        
        if any(kw in n for kw in ndt_keywords):
            # Check if it also contains equipment keywords (e.g. "MT 장비")
            if any(kw in n for kw in equip_keywords):
                return False
            return True
            
        # 2. RT consumables (Films)
        rt_keywords = ['FILM', 'CARESTREAM', 'MX125', 'T200', 'AA400', 'HS800', 'IX100', 'AGFA', 'FUJI']
        if any(kw in n for kw in rt_keywords):
            if any(kw in n for kw in equip_keywords): # e.g. "RT 장비"
                return False
            return True

        # 3. Method-based logic
        if m in ['MT', 'PT']:
            if any(kw in n for kw in equip_keywords):
                return False
            return True # Default to True for chemicals in MT/PT

        # 4. If name contains MT/PT but no equip keywords, and method is empty
        # This helps in Inventory Status loop where method is empty
        if not m and (n.startswith('PT') or n.startswith('MT')) and len(n) <= 10:
             if not any(kw in n for kw in equip_keywords):
                 return True

        # 5. Default: If method is PAUT/UT/RT/PMI, most items are equipment
        return False

    def save_daily_usage_edits(self, df_idx, new_data):
        """Save edited daily usage and reconcile stock"""
        try:
            old_entry = self.daily_usage_df.loc[df_idx]
            
            # 1. Revert Old Stock (Delete existing auto-deductions before applying new ones)
            old_site = old_entry.get('Site', '')
            old_date = pd.to_datetime(old_entry.get('Date'))
            
            if not self.transactions_df.empty:
                trans_mask = (
                    (pd.to_datetime(self.transactions_df['Date'], errors='coerce').dt.normalize() == pd.to_datetime(old_date).normalize()) &
                    (self.transactions_df['Site'] == old_site) &
                    (self.transactions_df['Type'] == 'OUT') &
                    (self.transactions_df['Note'].str.contains(f"{old_site} 현장 사용", na=False))
                )
                self.transactions_df = self.transactions_df[~trans_mask]

            # 2. Update entry in DF
            # [FIX] Handle column name aliases (Date/날짜, Site/현장) to ensure existing columns are updated
            site_col = 'Site' if 'Site' in self.daily_usage_df.columns else '현장'
            date_col = 'Date' if 'Date' in self.daily_usage_df.columns else '날짜'
            
            # [NEW] Pre-collect existing column names to avoid duplicate-like column creation
            existing_cols = self.daily_usage_df.columns
            
            for k, v in new_data.items():
                target_key = k
                if k == 'Site': target_key = site_col
                elif k == 'Date': target_key = date_col
                
                # [FIX] For NDT and RTK fields, ensure they match normalized column names in the DF
                if k.startswith('NDT_') or k.startswith('RTK_'):
                    # Search for case-insensitive and space-insensitive match in existing columns
                    k_norm = k.replace(' ', '').upper()
                    for ex_col in existing_cols:
                        if str(ex_col).replace(' ', '').upper() == k_norm:
                            target_key = ex_col
                            break
                
                # Comprehensive list of columns that MUST be numeric
                numeric_cols = ['Usage', '수량', '검사량', '단가', '출장비', '일식', '검사비', 'FilmCount', 'OT', 'OT금액']
                is_numeric = k.startswith('NDT_') or k.startswith('RTK_') or k in numeric_cols or any(f"OT{i}" == k or f"OT금액{i}" == k for i in range(1, 11))
                
                # Use safe setter to avoid all dtype crashes
                self._safe_set_daily_df(df_idx, target_key, v, is_numeric=(is_numeric or k == 'MaterialID'))

            # Ensure Usage is consistently updated from Inspection Amount (검사량)
            usage_col = 'Usage' if 'Usage' in self.daily_usage_df.columns else '수량'
            self._safe_set_daily_df(df_idx, usage_col, new_data.get('검사량', 0.0), is_numeric=True)
            if '검사량' in self.daily_usage_df.columns:
                self._safe_set_daily_df(df_idx, '검사량', new_data.get('검사량', 0.0), is_numeric=True)
            
            # [FIX] Initialize new_mat_display for stock reconciliation logic
            new_mat_display = new_data.get('Material', '')
            if not new_mat_display:
                new_mat_display = self.get_material_display_name(new_data.get('MaterialID', ''))
            
            # 3. Apply New Deduction (Selective Auto-deduction)
            new_date = pd.to_datetime(new_data['Date'])
            new_site = new_data['Site']
            # [NEW] PAUT 및 장비류는 재고 차감에서 제외
            method_raw = new_data.get('검사방법', '')
            method = str(method_raw).strip().upper()
            is_excluded = (method == 'PAUT')
            
            if not is_excluded:
                # Resolve or Register MaterialID
                new_mat_id = ""
                if new_mat_display:
                    for _, row in self.materials_df.iterrows():
                        if self.get_material_display_name(row['MaterialID']) == new_mat_display:
                            new_mat_id = row['MaterialID']
                            if row.get('Active', 1) == 0:
                                self.materials_df.loc[self.materials_df['MaterialID'] == new_mat_id, 'Active'] = 1
                            break
                    
                    if not new_mat_id and new_mat_display:
                        if self._is_consumable_material(new_mat_display, method):
                            new_mat_id = self.register_new_material(new_mat_display, warehouse='현장', 규격='자동등록')
                        else:
                            new_mat_id = new_mat_display
                
                new_qty = float(new_data.get('검사량', 0))
                new_note_pattern = f'{new_site} 현장 사용 (자동 차감)'
                
                workers_names = [new_data.get(f'User{i}' if i > 1 else 'User', '').strip() for i in range(1, 11)]
                all_workers = ", ".join([n for n in workers_names if n])
                
                if new_mat_id and new_qty > 0:
                    if self._is_consumable_material(new_mat_display, method):
                        mat_info = self.get_material_info(new_mat_id)
                        full_item_name = str(mat_info.get('품목명', '')).replace(' ', '').upper()
                        # PT/MT 약품 부모 항목 중복 차감 방지
                        if full_item_name not in ["PT약품", "MT약품", "NDT약품"]:
                            self._create_manual_stock_transaction(new_date, new_mat_id, 'OUT', new_qty, new_site, all_workers, new_note_pattern)
                
                # NDT 약품(소모품) 일괄 정산
                ndt_data = {
                    name: float(new_data.get(f'NDT_{"".join(name.split())}', 0))
                    for name in ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
                }
                self._auto_reconcile_and_register_ndt(new_date, new_site, ndt_data, all_workers, new_data.get('회사코드', ''))

            # 4. Save & Refresh
            if self.save_data():
                self.update_daily_usage_view()
                self.refresh_inquiry_filters()
                self.update_stock_view()
                self.update_transaction_view()
                self.update_material_combo() # Refresh history suggestions
                return True
            return False
        except Exception as e:
            messagebox.showerror("수정 오류", f"기록 수정 중 오류가 발생했습니다: {e}")
            return False
                
    def export_daily_usage_history(self):
        """Export current filtered daily usage history to Excel"""
        try:
            # Get selected items only, or all if none selected
            selected_items = self.daily_usage_tree.selection()
            if not selected_items:
                # If no items selected, export all visible items
                items_to_export = self.daily_usage_tree.get_children()
            else:
                # Export only selected items
                items_to_export = selected_items
            
            data = []
            for item in items_to_export:
                data.append(self.daily_usage_tree.item(item, 'values'))
            
            if not data:
                messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
                return
            
            # Export columns currently visible in the Treeview
            display_cols = self.daily_usage_tree['displaycolumns']
            if not display_cols or display_cols == ('#all'):
                selected_cols = list(self.daily_usage_tree['columns'])
            else:
                selected_cols = list(display_cols)
            
            # Map selected column names to their indices in the Treeview values
            all_cols = list(self.daily_usage_tree['columns'])
            col_indices = [all_cols.index(col) for col in selected_cols]
            
            # Reconstruct data
            filtered_data = []
            for item_values in data:
                row = [item_values[i] for i in col_indices if i < len(item_values)]
                filtered_data.append(row)
            
            filename = f"일일사용내역_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=filename, filetypes=[("Excel files", "*.xlsx")])
            
            if save_path:
                cols = list(self.daily_usage_tree['columns'])
                col_to_idx = {col: i for i, col in enumerate(cols)}
                
                # Rebuild data from 'data' (which has ALL columns including hidden ones)
                # and swap summarized '작업자' with full list from '(Full작업자)'
                final_filtered_data = []
                for item_values in data:
                    row_list = list(item_values)
                    
                    # Perform the swap if both columns exist
                    if '작업자' in cols and '(Full작업자)' in col_to_idx:
                        full_idx = col_to_idx['(Full작업자)']
                        if full_idx < len(row_list):
                            full_val = row_list[full_idx]
                            if full_val:
                                # Find index of '작업자' in ALL columns to swap in row_list
                                idx_worker = cols.index('작업자')
                                row_list[idx_worker] = full_val
                    
                    # Extract only the selected display columns for final export
                    final_row = [row_list[i] for i in col_indices if i < len(row_list)]
                    final_filtered_data.append(final_row)
                
                df = pd.DataFrame(final_filtered_data, columns=selected_cols)
                df = self.clean_df_export(df)
                self.save_df_to_excel_autofit(df, save_path, "일일사용내역")
                messagebox.showinfo("완료", f"데이터가 엑셀로 저장되었습니다.\n{save_path}")
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 내보내기 중 오류가 발생했습니다: {e}")

    def show_column_visibility_dialog(self, tree=None):
        """Open dialog to manually show/hide columns in the specified tree view"""
        if tree is None:
            tree = self.daily_usage_tree
            
        all_cols = list(tree['columns'])
        
        # Mandatory columns identification
        # NDT/RTK columns are at the end, workers in middle, basic info at start
        selectable_cols = [c for c in all_cols if c != '(Full작업자)']
        
        # [FIX] Use stored manual preferences instead of current filtered displaycolumns
        # so that hidden-by-data columns still appear checked if user wants them.
        if tree == self.daily_usage_tree:
            active_cols = getattr(self, 'manual_visible_cols', [])
            if not active_cols:
                # Default to all non-data columns if no preference saved
                active_cols = [c for c in all_cols if c != '(Full작업자)']
        else:
            active_cols = tree['displaycolumns']
            if not active_cols or active_cols == ('#all'):
                 active_cols = all_cols
        
        dialog = ColumnSelectionDialog(self.root, selectable_cols, title="표시 컬럼 관리")
        # Overwrite vars with current visibility state
        for col, var in dialog.vars.items():
            var.set(col in active_cols)
            
        dialog.wait_window()
        
        if dialog.result is not None:
            # Result set from dialog
            final_selection = list(dialog.result)
            
            # Reconstruct final display based on user result + mandatory ones
            mandatory_lookup = {
                self.daily_usage_tree: ['날짜', '업체명', '현장', '작업자', '검사품명', '수량', '단위', '단가', '검사비', 'OT시간', 'OT금액', '비고'],
                self.inout_tree: ['날짜', '구분', '품목명']
            }
            
            mandatory_cols = mandatory_lookup.get(tree, ['날짜'])
            for mc in mandatory_cols:
                if mc in all_cols and mc not in final_selection:
                    final_selection.append(mc)
            
            # Re-order to match original tree columns
            sorted_selection = [c for c in all_cols if c in final_selection]
            
            tree['displaycolumns'] = sorted_selection
            
            # Save configuration to the appropriate key
            if tree == self.daily_usage_tree:
                self.manual_visible_cols = sorted_selection
                if not hasattr(self, 'tab_config'): self.tab_config = {}
                self.tab_config['daily_usage_visible_cols'] = sorted_selection
            elif tree == self.inout_tree:
                if not hasattr(self, 'tab_config'): self.tab_config = {}
                self.tab_config['inout_visible_cols'] = sorted_selection
                
            self.save_tab_config()
            
            # Immediate refresh to apply selection
            if tree == self.daily_usage_tree:
                self.update_daily_usage_view()
            elif tree == self.inout_tree:
                self.update_transaction_view()

    def export_all_daily_usage(self):
        """Export all daily usage records to Excel"""
        try:
            if self.daily_usage_df.empty:
                messagebox.showinfo("알림", "기록된 데이터가 없습니다.")
                return
            
            # Export all columns from the tree structure
            selected_cols = list(self.daily_usage_tree['columns'])
            
            filename = f"전체사용기록_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=filename, filetypes=[("Excel files", "*.xlsx")])
            
            if save_path:
                # Ensure we only try to export columns that exist in the dataframe
                valid_cols = [c for c in selected_cols if c in self.daily_usage_df.columns]
                # If dataframe has extra columns not in tree (like raw data), include them? 
                # Better to stick to what the user sees/defines in tree interface + critical data
                # Actually, export_all usually implies backing up the raw dataframe. 
                # But previous logic used tree columns. Let's stick to tree columns if they exist in DF.
                
                # Fallback: if tree columns are not in DF (e.g. calculated columns), we might have issues.
                # The daily_usage_df has raw data. The tree has formatted data.
                # daily_usage_df columns: Date, Site, MaterialID, Usage, Note, etc. + RTK_..., NDT_...
                # Tree columns: 날짜, 현장, 작업자... OT합계...
                
                # If we want "All Daily Usage", we should probably export the RAW dataframe for backup purposes, 
                # OR the formatted view for reporting.
                # The previous code seemed to try to export based on tree columns but mapped from DF?
                # "export_df = self.daily_usage_df[selected_cols].copy()"
                # This implies 'selected_cols' MUST exist in daily_usage_df.
                # Let's check daily_usage_df columns again.
                # It has 'Date', 'Site', 'MaterialID'...
                # It does NOT have '날짜', '현장' (Korean names).
                # So the previous code 'self.daily_usage_df[selected_cols]' would have FAILED if selected_cols came from tree columns!
                
                # WAIT. The user said "remove column selection". 
                # The previous working code (before my changes) probably did:
                # export_df = self.daily_usage_df.copy()
                # Let's revert to a safe "Backup" style export for "All Data".
                
                export_df = self.daily_usage_df.copy()
                export_df = self.clean_df_export(export_df)
                self.save_df_to_excel_autofit(export_df, save_path, "전체사용기록")
                
                export_df = self.clean_df_export(export_df)
                self.save_df_to_excel_autofit(export_df, save_path, "전체기록")
                messagebox.showinfo("완료", f"전체 기록이 엑셀로 저장되었습니다.\n{save_path}")
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 내보내기 중 오류가 발생했습니다: {e}")


    def setup_keyboard_shortcuts(self):
        """Setup keyboard shortcuts for navigation"""
        # Ctrl+Tab to switch between notebook tabs (forward)
        self.root.bind('<Control-Tab>', self.next_tab)
        # Ctrl+Shift+Tab to switch between notebook tabs (backward)
        self.root.bind('<Control-Shift-Tab>', self.prev_tab)
        
        # Alt+숫자 for direct tab access
        self.root.bind('<Alt-Key-1>', lambda e: self.notebook.select(0))
        self.root.bind('<Alt-Key-2>', lambda e: self.notebook.select(1))
        self.root.bind('<Alt-Key-3>', lambda e: self.notebook.select(2))
        self.root.bind('<Alt-Key-4>', lambda e: self.notebook.select(3))
        self.root.bind('<Alt-Key-5>', lambda e: self.notebook.select(4))
        self.root.bind('<Alt-Key-6>', lambda e: self.notebook.select(5))
        
        # Right-click on notebook for tab reordering
        # Tab interactions removed here, kept in create_widgets
    
    def next_tab(self, event=None):
        """Switch to next tab"""
        current = self.notebook.index(self.notebook.select())
        total = self.notebook.index('end')
        next_tab = (current + 1) % total
        self.notebook.select(next_tab)
        return 'break'
    
    def prev_tab(self, event=None):
        """Switch to previous tab"""
        current = self.notebook.index(self.notebook.select())
        total = self.notebook.index('end')
        prev_tab = (current - 1) % total
        self.notebook.select(prev_tab)
        return 'break'
    
    def show_tab_context_menu(self, event):
        """Show context menu for tab reordering"""
        # Identify which tab was clicked
        try:
            clicked_tab = self.notebook.index(f"@{event.x},{event.y}")
            self.notebook.select(clicked_tab)
            
            # Create context menu
            context_menu = tk.Menu(self.root, tearoff=0)
            
            # Only show "Move Left" if not the first tab
            if clicked_tab > 0:
                context_menu.add_command(label="← 탭 왼쪽으로 이동", 
                                        command=lambda: self.move_tab_left(clicked_tab))
            
            # Only show "Move Right" if not the last tab
            if clicked_tab < self.notebook.index('end') - 1:
                context_menu.add_command(label="탭 오른쪽으로 이동 →", 
                                        command=lambda: self.move_tab_right(clicked_tab))
            
            # Show menu at cursor position
            context_menu.post(event.x_root, event.y_root)
        except:
            pass
    
    def move_tab_left(self, tab_index):
        """Move tab one position to the left"""
        if tab_index > 0:
            # Get tab info
            tab = self.notebook.tabs()[tab_index]
            text = self.notebook.tab(tab_index, "text")
            
            # Remove and reinsert at new position
            self.notebook.insert(tab_index - 1, tab, text=text)
            self.notebook.select(tab_index - 1)
            
            # Save configuration immediately
            self.save_tab_config()
    
    def move_tab_right(self, tab_index):
        """Move tab one position to the right"""
        if tab_index < self.notebook.index('end') - 1:
            # Get tab info
            tab = self.notebook.tabs()[tab_index]
            text = self.notebook.tab(tab_index, "text")
            
            # Remove and reinsert at new position
            self.notebook.insert(tab_index + 2, tab, text=text)
            self.notebook.select(tab_index + 1)
            
            # Save configuration immediately
            self.save_tab_config()

    def on_tab_drag_start(self, event):
        """Start tab dragging by identifying the tab under the cursor"""
        try:
            # Check if click is on a tab tab
            clicked_tab = self.notebook.identify(event.x, event.y)
            if clicked_tab == "label":
                # Find which index this is
                self._drag_start_index = self.notebook.index(f"@{event.x},{event.y}")
                self._current_drag_index = self._drag_start_index
            else:
                self._drag_start_index = None
        except:
            self._drag_start_index = None

    def on_tab_drag(self, event):
        """Handle visual swapping of tabs during drag"""
        if not hasattr(self, '_drag_start_index') or self._drag_start_index is None:
            return

        try:
            # Find the index of the tab currently under the cursor
            target_index = self.notebook.index(f"@{event.x},{event.y}")
            
            if target_index != self._current_drag_index:
                # Get the widget of the tab we are dragging
                tab_widget = self.notebook.tabs()[self._current_drag_index]
                tab_text = self.notebook.tab(self._current_drag_index, "text")
                
                # Use insert to move tab widget. 
                # Note: insert(pos, widget) handles the reordering logic in Notebook
                self.notebook.insert(target_index, tab_widget, text=tab_text)
                
                self._current_drag_index = target_index
                self.notebook.select(target_index)
        except:
            pass

    def normalize_site_name(self, name):
        """[ROBUST] Normalize site names: strip whitespace, unify ALL hyphen types, collapse multiple spaces."""
        if name is None or pd.isna(name): return ""
        import re
        s = str(name).strip()
        # Unify various types of hyphens/dashes: -, –, —, －
        s = re.sub(r'[\u002D\u2013\u2014\uFF0D]', '-', s)
        # Unify hyphens by removing spaces around them: "Site - A" -> "Site-A"
        s = re.sub(r'\s*-\s*', '-', s)
        # Collapse any multiple internal spaces
        s = " ".join(s.split())
        return s

    def on_tab_drag_end(self, event):
        """Finalize tab order and save configuration"""
        if hasattr(self, '_drag_start_index') and self._drag_start_index is not None:
            if self._current_drag_index != self._drag_start_index:
                print(f"Tab reordered: {self._drag_start_index} -> {self._current_drag_index}")
                # [FIX] Force save on manual drag end to ensure order is persisted
                self.save_tab_config(force=True)
            
            self._drag_start_index = None
            self._current_drag_index = None
    
    def on_tab_changed(self, event=None):
        """Handle tab selection change event"""
        try:
            # 1. Save configuration when tab changes (respects is_ready via save_tab_config)
            self.save_tab_config()

            # 1-1. 탭 이동 시 실제 데이터(Excel)도 자동 저장
            if getattr(self, 'is_ready', False):
                self.save_data()
            
            # 2. Handle specific tab UI adjustments
            current_tab = self.notebook.select()
            if not current_tab:
                return
                
            # Convert widget path to index if needed
            try:
                current_tab_idx = self.notebook.index("current")
                tab_text = self.notebook.tab(current_tab_idx, "text")
            except:
                tab_text = ""

            # Check for Daily Usage tab
            if (hasattr(self, 'tab_daily_usage') and str(current_tab) == str(self.tab_daily_usage)) or \
               tab_text == '현장별 일일 사용량 기입':
                
                print("Daily usage tab selected - ensuring visibility")
                # Force multiple updates when tab is selected
                self.refresh_inquiry_filters()
                self.update_daily_usage_view() # [NEW] Trigger auto-hiding logic
                self.root.after(50, self._ensure_daily_usage_sash_visibility)
                self.root.after(200, self._ensure_daily_usage_sash_visibility)
                self.root.after(400, self._ensure_canvas_scroll_region)
                
                # Also ensure the inner history sash is visible
                self.root.after(150, self._ensure_sash_visible)
            
            elif tab_text == '월별 집계':
                print("Monthly usage tab selected - refreshing view")
                self.refresh_inquiry_filters()
                self.update_monthly_usage_view()
            elif tab_text == '공사실행예산서':
                print("Construction budget tab selected - refreshing view")
                self.update_budget_site_view()
            elif tab_text == '입출고 관리':
                print("In/Out Management tab selected - refreshing history")
                self.refresh_inout_history()
        except Exception as e:
            print(f"Error in tab change handler: {e}")
    
    def auto_save_to_list(self, event, combobox, data_list, config_key):
        """Helper to auto-save new entry from combobox to a list and update all related UI"""
        new_val = combobox.get().strip()
        if not new_val:
            return
            
        if new_val not in data_list:
            data_list.append(new_val)
            data_list.sort()
            
            # Trigger app-wide update
            self.refresh_ui_for_list_change(config_key)
            self.save_tab_config()


    def auto_save_worktime(self, event, entry, config_key):
        """Helper to auto-save worktime values and support copy functionality"""
        worktime_value = entry.get().strip()
        if not worktime_value:
            return
            
        # Initialize worktimes list if not exists
        if not hasattr(self, 'worktimes'):
            self.worktimes = []
            
        # Check if this value already exists
        if worktime_value not in self.worktimes:
            self.worktimes.append(worktime_value)
            self.worktimes.sort()
            
            # Save to config and update UI
            try:
                self.refresh_ui_for_list_change('worktimes')
                self.save_tab_config()
            except Exception as e:
                print(f"Failed to save worktime config: {e}")
        
        # Calculate and update corresponding OT field
        self.calculate_and_update_ot(worktime_value, entry)
        
        # Add copy functionality - select all text on focus and bind Ctrl+C (only once)
        entry.select_range(0, tk.END)
        if not hasattr(entry, '_copy_bound'):
            entry.bind('<Control-c>', lambda e: self.copy_worktime(entry))
            entry._copy_bound = True
        
        # Store current value for copy functionality
        entry.last_value = worktime_value

    def calculate_and_update_ot(self, worktime_value, worktime_entry):
        """Calculate OT amount based on worktime and update corresponding OT field"""
        try:
            current_date = self.ent_daily_date.get_date()
            ot_hours, amount = self._calculate_ot_from_worktime(worktime_value, current_date)
            
            ot_entry = self.get_corresponding_ot_field(worktime_entry)
            if ot_entry:
                if amount > 0:
                    ot_entry.delete(0, tk.END)
                    ot_entry.insert(0, f"{amount:,}")
                else:
                    ot_entry.delete(0, tk.END)
                    # If it was cleared, but hours > 0 (very small amount?), maybe set to 0
                    if ot_hours > 0: ot_entry.insert(0, "0")
        except Exception as e:
            print(f"Error calculating OT: {e}")

    def get_corresponding_ot_field(self, worktime_entry):
        """Get corresponding OT field for a worktime entry"""
        # Map worktime entries to OT entries
        worktime_to_ot = {
            'ent_worktime1': 'ent_ot1', 'ent_worktime2': 'ent_ot2', 
            'ent_worktime3': 'ent_ot3', 'ent_worktime4': 'ent_ot4',
            'ent_worktime5': 'ent_ot5', 'ent_worktime6': 'ent_ot6',
            'ent_worktime7': 'ent_ot7', 'ent_worktime8': 'ent_ot8',
            'ent_worktime9': 'ent_ot9', 'ent_worktime10': 'ent_ot10'
        }
            
        # Find worktime entry name by checking object IDs
        worktime_id = id(worktime_entry)
            
        worktime_attrs = [f'ent_worktime{i}' for i in range(1, 11)]
        for attr_name in worktime_attrs:
            if hasattr(self, attr_name):
                attr_value = getattr(self, attr_name)
                if id(attr_value) == worktime_id:
                    ot_attr_name = worktime_to_ot.get(attr_name)
                    if ot_attr_name and hasattr(self, ot_attr_name):
                        return getattr(self, ot_attr_name)
                    break
            
        return None

    def copy_worktime(self, entry):
        """Copy worktime value to clipboard"""
        try:
            worktime_value = entry.get().strip()
            if worktime_value:
                self.root.clipboard_clear()
                self.root.clipboard_append(worktime_value)
                # Show brief feedback
                entry.delete(0, tk.END)
                entry.insert(0, worktime_value)
        except Exception as e:
            pass  # Silently handle clipboard errors

    def auto_save_ot(self, event, entry, config_key):
        """Helper to auto-save OT values and support copy functionality"""
        ot_value = entry.get().strip()
        if not ot_value:
            return
                
        # Initialize ot_times list if not exists
        if not hasattr(self, 'ot_times'):
            self.ot_times = []
            
        # Calculate OT amount for any input
        calculated_amount = self.calculate_ot_amount(ot_value)
        if calculated_amount:
            # Update entry with calculated amount
            entry.delete(0, tk.END)
            display_value = f"{ot_value} ({calculated_amount:,}원)"
            entry.insert(0, display_value)
            ot_value = display_value
            print(f"OT calculated: {ot_value}")  # Debug print
            
        # Check if this value already exists
        if ot_value not in self.ot_times:
            self.ot_times.append(ot_value)
            self.ot_times.sort()
            
        # Save to config with error handling - only save when list changes
        try:
            self.save_tab_config()
        except Exception as e:
            print(f"Failed to save OT config: {e}")
            
        # Add copy functionality - select all text on focus and bind Ctrl+C (only once)
        entry.select_range(0, tk.END)
        if not hasattr(entry, '_copy_bound'):
            entry.bind('<Control-c>', lambda e: self.copy_ot(entry))
            entry._copy_bound = True
            
        # Store current value for copy functionality
        entry.last_value = ot_value

    def calculate_ot_amount(self, ot_value):
        """Calculate OT amount based on time and rates, or just parse if already an amount"""
        try:
            if not ot_value or not str(ot_value).strip(): return 0
            val = str(ot_value).strip().replace(',', '')
            
            # If it's already just a large number, it's the amount
            if val.isdigit() and int(val) > 100:
                return int(val)
            
            # If it has (N원) format
            if '(' in val and '원)' in val:
                try:
                    return int(val.split('(')[1].split('원')[0].replace(',', ''))
                except: pass

            hours = self._parse_ot_hours(val)
            if hours <= 0: return 0

            # (Rest of simple duration fallback)
            evening_hours = min(hours, 4)
            night_hours = max(0, hours - 4)
            return int(evening_hours * 4000 + night_hours * 5000)
        except Exception as e:
            return 0

    def _parse_ot_hours(self, ot_value):
        """Helper to extract numeric OT hours using regex for maximum robustness"""
        import re
        try:
            if not ot_value or not str(ot_value).strip(): return 0
            val = str(ot_value).strip().replace(' ', '').replace('익일', '')
            
            # If it's just a large number (>100), assume it's an amount, not hours
            if val.replace(',', '').isdigit() and int(val.replace(',', '')) > 100:
                return 0.0

            # 1. Check for "N시간"
            dur_match = re.search(r'(\d+\.?\d*)\s*(시간|hr|h)', val)
            if dur_match:
                return float(dur_match.group(1))

            # 2. Check for time range "18:00~22:00"
            range_match = re.search(r'(\d{1,2}):(\d{1,2})[-~](\d{1,2}):(\d{1,2})', val)
            if range_match:
                h1, m1, h2, m2 = map(int, range_match.groups())
                if h2 < h1: h2 += 24
                return (h2 * 60 + m2 - (h1 * 60 + m1)) / 60

            # 3. Check for simple ':' format "2:30"
            colon_match = re.search(r'^(\d{1,2}):(\d{1,2})$', val)
            if colon_match:
                h, m = map(int, colon_match.groups())
                return h + (m / 60)

            # 4. Fallback to just extracting the first small float/int found
            num_match = re.search(r'(\d+\.?\d*)', val)
            if num_match:
                v = float(num_match.group(1))
                if v <= 24: return v # Reasonable hour count
            
            return 0
        except:
            return 0

    def _calculate_split_ot_hours(self, ot_value, date_val=None):
        """Split OT hours into Day window (18-22), Night window (22-24), and Holiday window for weekends/Friday dawn"""
        import re
        import pandas as pd
        try:
            if not ot_value or not str(ot_value).strip(): return 0.0, 0.0, 0.0
            val = str(ot_value).strip().replace(' ', '')
            
            total_hours = self._parse_ot_hours(val)
            if total_hours <= 0: return 0.0, 0.0, 0.0

            # Default start at 18:00 if no range
            start_hour = 18
            range_match = re.search(r'(\d{1,2}):(\d{1,2})[-~](\d{1,2}):(\d{1,2})', val)
            if range_match:
                start_hour = int(range_match.group(1))
            
            # Handle overnight logic (e.g., 18:00~01:00)
            current_time = float(start_hour)
            remaining = total_hours
            day_hours = 0.0
            night_hours = 0.0
            holiday_hours = 0.0
            
            # Check if this is weekend (Sat/Sun) or Friday going into Saturday
            is_friday = False
            is_weekend = False
            if date_val is not None:
                try:
                    dt = pd.to_datetime(date_val)
                    is_friday = (dt.weekday() == 4)
                    is_weekend = (dt.weekday() >= 5) # Sat=5, Sun=6
                except:
                    pass
            
            # [FIXED] If it's Saturday or Sunday, all overtime is holiday work
            if is_weekend:
                return 0.0, 0.0, float(total_hours)

            # Simulate hour by hour (or portion by portion)
            while remaining > 0:
                # 18:00 ~ 22:00 구간은 연장근무 (day_hours)
                if 18 <= current_time < 22:
                    can_take = 22 - current_time
                    taken = min(remaining, can_take)
                    day_hours += taken
                    remaining -= taken
                    current_time += taken
                # 22:00 ~ 24:00 구간은 야간근무 (night_hours)
                elif 22 <= current_time < 24:
                    can_take = 24 - current_time
                    taken = min(remaining, can_take)
                    night_hours += taken
                    remaining -= taken
                    current_time += taken
                # 24:00 ~ (익일) 구간
                elif current_time >= 24:
                    can_take = remaining # 끝까지 처리
                    taken = min(remaining, can_take)
                    if is_friday:
                        holiday_hours += taken
                    else:
                        night_hours += taken
                    remaining -= taken
                    current_time += taken
                # 18:00 이전 시간이 혹시 OT로 입력되었다면 상황에 맞게 처리 (기본 시뮬레이션에서는 18시까지 대기하는 것으로 가정)
                else: 
                     current_time = 18.0
            
            return day_hours, night_hours, holiday_hours
        except:
            return 0.0, 0.0, 0.0

    def copy_ot(self, entry):
        """Copy OT value to clipboard"""
        try:
            ot_value = entry.get().strip()
            if ot_value:
                self.root.clipboard_clear()
                self.root.clipboard_append(ot_value)
                # Show brief feedback
                entry.delete(0, tk.END)
                entry.insert(0, ot_value)
        except Exception as e:
            pass  # Silently handle clipboard errors


    def save_tab_config(self, force=False):
        """Save current tab configuration to memory and disk"""
        try:
            # [STABILITY] Skip saving if app is not fully ready (prevents overwriting saved tab index during init)
            if not getattr(self, 'is_ready', False) and not force:
                return

            # Force update to ensure all coordinates and sizes are accurate
            self.root.update_idletasks()
            
            # Initialize or keep existing tab_config
            if not hasattr(self, 'tab_config'):
                self.tab_config = {}
                
            # Update with current core values
            current_tab_idx = 0
            current_tab_text = ""
            try:
                current_tab_widget = self.notebook.select()
                current_tab_idx = self.notebook.index(current_tab_widget)
                current_tab_text = self.notebook.tab(current_tab_widget, "text")
            except:
                current_tab_idx = self.tab_config.get('selected_tab', 0)
                current_tab_text = self.tab_config.get('selected_tab_text', "")

            # Get current tab order
            tab_order = []
            for tab_id in self.notebook.tabs():
                tab_order.append(self.notebook.tab(tab_id, "text"))

            self.tab_config.update({
                'selected_tab': current_tab_idx,
                'selected_tab_text': current_tab_text,
                'tab_order': tab_order,
                'sites': self.sites,
                'daily_units': self.daily_units,
                'users': getattr(self, 'users', []),
                'companies': getattr(self, 'companies', []),
                'warehouses': getattr(self, 'warehouses', []),
                'equipments': getattr(self, 'equipments', []),
                'vehicles': getattr(self, 'vehicles', []),
                'materials': getattr(self, 'carestream_films', []),
                'worktimes': getattr(self, 'worktimes', []),
                'ot_times': getattr(self, 'ot_times', []),
                'layout_locked': getattr(self, 'layout_locked', False),
                'resolution_locked': getattr(self, 'resolution_locked', False),
                'locked_width': getattr(self, 'locked_width', 1200),
                'locked_height': getattr(self, 'locked_height', 800),
                'daily_usage_sash_locked': self.daily_usage_sash_locked,
                'daily_usage_sash_pos': self.daily_usage_paned.sashpos(0) if hasattr(self, 'daily_usage_paned') else None,
                'daily_history_sash_pos': self.daily_history_paned.sashpos(0) if hasattr(self, 'daily_history_paned') else None,
                'entry_inner_frame_height': self.entry_inner_frame.winfo_height() if hasattr(self, 'entry_inner_frame') else None,
                'history_visible_cols': getattr(self, 'manual_visible_cols', []),
                'monthly_visible_cols': getattr(self, 'monthly_visible_cols', []),
                'budget_view_visible_cols': getattr(self, 'budget_view_visible_cols', []),
                'budget_view_heading_aliases': getattr(self, 'budget_view_heading_aliases', {}),
                'budget_view_custom_columns': getattr(self, 'budget_view_custom_columns', []),
                'window_state': self.root.state(),
                'window_width': self.root.winfo_width(),
                'window_height': self.root.winfo_height()
            })
            
            # Save current stock column widths
            self.tab_config['stock_col_widths'] = {}
            if hasattr(self, 'stock_tree'):
                for col in self.stock_tree['columns']:
                    self.tab_config['stock_col_widths'][col] = self.stock_tree.column(col, 'width')

            # Save in/out history column widths
            self.tab_config['inout_col_widths'] = {}
            if hasattr(self, 'inout_tree'):
                for col in self.inout_tree['columns']:
                    self.tab_config['inout_col_widths'][col] = self.inout_tree.column(col, 'width')

            # Save daily usage column widths
            self.tab_config['daily_usage_col_widths'] = {}
            if hasattr(self, 'daily_usage_tree'):
                for col in self.daily_usage_tree['columns']:
                    self.tab_config['daily_usage_col_widths'][col] = self.daily_usage_tree.column(col, 'width')
            
            # Save monthly usage column widths
            self.tab_config['monthly_usage_col_widths'] = {}
            if hasattr(self, 'monthly_usage_tree'):
                for col in self.monthly_usage_tree['columns']:
                    self.tab_config['monthly_usage_col_widths'][col] = self.monthly_usage_tree.column(col, 'width')
            
            # [NEW] Save summary table column widths
            self.tab_config['site_summary_col_widths'] = {}
            if hasattr(self, 'site_summary_tree'):
                for col in self.site_summary_tree['columns']:
                    self.tab_config['site_summary_col_widths'][col] = self.site_summary_tree.column(col, 'width')
            
            self.tab_config['worker_summary_col_widths'] = {}
            if hasattr(self, 'worker_summary_tree'):
                for col in self.worker_summary_tree['columns']:
                    self.tab_config['worker_summary_col_widths'][col] = self.worker_summary_tree.column(col, 'width')
            
            # [NEW] Save report column widths
            self.tab_config['report_col_widths'] = {}
            if hasattr(self, 'report_tree'):
                for col in self.report_tree['columns']:
                    self.tab_config['report_col_widths'][col] = self.report_tree.column(col, 'width')

            # Save budget site performance column widths
            self.tab_config['budget_view_col_widths'] = {}
            if hasattr(self, 'budget_view_tree'):
                for col in self.budget_view_tree['columns']:
                    self.tab_config['budget_view_col_widths'][col] = self.budget_view_tree.column(col, 'width')
            
            # Save tab order
            self.tab_config['tab_order'] = []
            for tab in self.notebook.tabs():
                self.tab_config['tab_order'].append(self.notebook.tab(tab, "text"))
                
            # [NEW] Save equipment list
            self.tab_config['equipments'] = self.equipments
            
            # Save Draggable items
            # Prepare in-memory geometries for merging (capture current states)
            current_geometries = {}
            
            for key, widget in self.draggable_items.items():
                # [LAYOUT FIX] 핵심 박스는 좌표 저장 제외
                # 재시작 시 기본 grid 배치로 복원하여 겹침/잘림 방지
                if key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
                    continue

                manager = widget.winfo_manager()
                # Capture everything that is visible and managed
                if manager in ['place', 'grid']:
                    current_geometries[key] = {
                        'x': widget.winfo_x(),
                        'y': widget.winfo_y(),
                        'width': widget.winfo_width(),
                        'height': widget.winfo_height(),
                        'hidden': False
                    }
                    
                    if hasattr(widget, '_label_widget'):
                        current_geometries[key]['custom_label'] = widget._label_widget.cget('text')
                    if hasattr(widget, '_manage_list_key') and widget._manage_list_key:
                        current_geometries[key]['manage_list_key'] = widget._manage_list_key
                    if key.startswith('clone_'):
                        current_geometries[key]['is_clone'] = True
                        current_geometries[key]['widget_class_name'] = widget._widget_class.__name__
                        saved_kwargs = widget._widget_kwargs.copy()
                        if 'values' in saved_kwargs: del saved_kwargs['values']
                        current_geometries[key]['widget_kwargs'] = saved_kwargs
                    if key.startswith('memo_'):
                        current_geometries[key]['text'] = self.memos[key]['text_widget'].get('1.0', 'end-1c')
                        current_geometries[key]['memo_title'] = self.memos[key]['title_entry'].get()
                    if key in self.checklists:
                        current_geometries[key]['checklist_title'] = self.checklists[key]['title_entry'].get()
                        items_data = []
                        for child in self.checklists[key]['item_frame'].winfo_children():
                            if hasattr(child, '_checklist_data'):
                                items_data.append({
                                    'text': child._checklist_data['entry'].get(),
                                    'checked': child._checklist_data['var'].get()
                                })
                        current_geometries[key]['checklist_items'] = items_data
                    if key in self.vehicle_inspections:
                        current_geometries[key]['vehicle_data'] = self.vehicle_inspections[key].get_data()
                elif manager == '': # Hidden or not managed
                    # Only mark as hidden if it was previously in place/grid (to avoid junk)
                    if key in self.tab_config.get('draggable_geometries', {}):
                        current_geometries[key] = {'hidden': True}

            # [DEFINITIVE FIX] Deep merge current_geometries into disk config
            final_config = {}
            if os.path.exists(self.config_path):
                try:
                    with open(self.config_path, 'r', encoding='utf-8') as f:
                        final_config = json.load(f)
                except: pass
            
            # Use disk's geometries as base, then update with WHAT WE JUST CAPTURED in memory
            disk_geos = final_config.get('draggable_geometries', {})

            # [LAYOUT FIX] 과거에 저장된 핵심 박스 좌표도 제거
            for core_key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
                if core_key in disk_geos:
                    del disk_geos[core_key]

            disk_geos.update(current_geometries)
            
            # Apply other tab state updates
            final_config.update(self.tab_config)
            final_config['draggable_geometries'] = disk_geos
            # [NEW] Save hidden_sites to config
            final_config['hidden_sites'] = list(getattr(self, 'hidden_sites', []))
            
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(final_config, f, ensure_ascii=False, indent=2)
            
            # Update in-memory copy as well
            self.tab_config = final_config
            
            print(f"Configuration saved. Locked: {self.tab_config.get('layout_locked')}, Tab: {self.tab_config.get('selected_tab')}")
        except Exception as e:
            print(f"Failed to save tab config: {e}")

    
    def load_tab_config(self):
        """Load and restore tab configuration"""
        try:
            if not hasattr(self, 'tab_config'):
                self.tab_config = {}
                
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    self.tab_config = json.load(f)
                    print(f"DEBUG: Loaded config raw: {self.tab_config}")
                
                config = self.tab_config
                
                # Restore lists - Respecting user deletions
                if 'daily_units' in config:
                    loaded = config['daily_units']
                    if loaded: # If user has a saved list (even if they deleted items)
                        self.daily_units = list(dict.fromkeys(loaded))
                    else: # Fallback only if the list is literally empty
                        self.daily_units = ['EA', 'CAN', 'SET', 'KG', 'M', '매', 'I/D', 'P,M,I/D', 'M,I/D', 'Point', 'Meter', 'Inch', 'Dia']
                
                # Apply high-level lock states immediately to internal variables
                self.layout_locked = config.get('layout_locked', False)
                self.resolution_locked = config.get('resolution_locked', False)
                self.daily_usage_sash_locked = config.get('daily_usage_sash_locked', False)
                self.locked_width = config.get('locked_width', 1200)
                self.locked_height = config.get('locked_height', 800)

                # Get current tab order
                current_order = []
                for tab in self.notebook.tabs():
                    tab_index = self.notebook.index(tab)
                    tab_text = self.notebook.tab(tab_index, "text")
                    current_order.append((tab_text, tab))
                
                # Create mapping from text to tab widget
                tab_map = {text: tab for text, tab in current_order}
                
                # Restore tab order if saved order exists
                saved_order = config.get('tab_order', [])
                if saved_order:
                    # [ROBUST] Reorder tabs according to saved order (relaxed length check)
                    for i, tab_text in enumerate(saved_order):
                        if tab_text in tab_map:
                            tab = tab_map[tab_text]
                            current_pos = self.notebook.index(tab)
                            if current_pos != i:
                                self.notebook.insert(i, tab, text=tab_text)
                
                # Restore selected tab
                selected_idx = config.get('selected_tab', 0)
                selected_text = config.get('selected_tab_text', "")
                
                tab_restored = False
                if selected_text:
                    # [ROBUST] Find tab by name instead of index
                    for i in range(self.notebook.index('end')):
                        if self.notebook.tab(i, "text") == selected_text:
                            try:
                                self.notebook.select(i)
                                tab_restored = True
                                break
                            except: pass
                
                if not tab_restored and 0 <= selected_idx < self.notebook.index('end'):
                    try:
                        self.notebook.select(selected_idx)
                    except: pass
                
                # Force update after selection so the tab is rendered and children computed
                self.root.update_idletasks()
                
                # Restore sites list - use in-place update
                self.sites[:] = config.get('sites', [])
                # If sites list is empty, try to populate from current daily_usage_df
                if not self.sites and not self.daily_usage_df.empty:
                    # [FIX] Collect and normalize sites in-place to maintain widget references and resolve crash
                    self.sites[:] = sorted(list(self.daily_usage_df['Site'].dropna().astype(str).apply(self.normalize_site_name).unique()))
                # [NEW] Restore hidden sites list and ensure requested defaults are present
                saved_hidden = config.get('hidden_sites', [])
                for s in ["초안", "롯데현장"]:
                    if s not in saved_hidden: saved_hidden.append(s)
                self.hidden_sites[:] = saved_hidden
                
                # Restore users list - use in-place update to preserve references
                self.users[:] = config.get('users', [])
                
                # [MIGRATION] Load and immediately migrate worktimes to ensure 익일 marker
                raw_worktimes = config.get('worktimes', [])
                self.worktimes = self._migrate_worktimes(raw_worktimes)
                
                # Restore vehicles, equipments, and preferred materials
                self.vehicles[:] = config.get('vehicles', [])
                self.equipments[:] = config.get('equipments', [])
                if 'materials' in config:
                    self.carestream_films[:] = config['materials']
                
                # [NEW] Restore companies list
                self.companies[:] = config.get('companies', [])
                if hasattr(self, 'cb_daily_company'):
                    self.cb_daily_company['values'] = sorted(self.companies)

                
                # Refresh all WorkerDataGroup dropdowns with the migrated list
                for i in range(1, 11):
                    group_attr = f'worker_group{i}'
                    if hasattr(self, group_attr):
                        getattr(self, group_attr).update_time_list(self.worktimes)

                self.ot_times = config.get('ot_times', [])

                # [CLEANUP] Clean up users list from shift markers
                import re
                cleaned_users = []
                for user in self.users:
                    if re.match(r"^\((주간|야간|휴일)\)$", user.strip()):
                        continue
                    match = re.match(r"\((주간|야간|휴일)\)\s*(.*)", user.strip())
                    if match:
                        actual_name = match.group(2).strip()
                        if actual_name and actual_name not in cleaned_users:
                            cleaned_users.append(actual_name)
                    elif user.strip() and user.strip() not in cleaned_users:
                        cleaned_users.append(user.strip())
                
                if len(cleaned_users) != len(self.users):
                    self.users[:] = cleaned_users
                    # Don't save yet, let the normal flow handle it
                
                # Restore history visibility if saved
                # [FIX] Mismatch between 'history_visible_cols' and 'daily_usage_visible_cols'
                self.manual_visible_cols = config.get('daily_usage_visible_cols', config.get('history_visible_cols', []))
                
                # [MIGRATION] If 'OT합계' is in saved cols, replace with 'OT시간', 'OT금액'
                if 'OT합계' in self.manual_visible_cols:
                    index = self.manual_visible_cols.index('OT합계')
                    self.manual_visible_cols.pop(index)
                    if 'OT시간' not in self.manual_visible_cols:
                        self.manual_visible_cols.insert(index, 'OT시간')
                    if 'OT금액' not in self.manual_visible_cols:
                        self.manual_visible_cols.insert(index + 1, 'OT금액')
                
                # [MIGRATION] Translate English names to Korean
                translations = {
                    'Date': '날짜', 'Site': '현장', 'User': '작업자', 'WorkTime': '작업시간',
                    'Equipment': '장비명', 'Method': '검사방법', 'Remark': '비고', 
                    'Note': '비고', 'EntryTime': '입력시간', 'MaterialName': '품목명',
                    '차량번호': '차량번호', '주행거리': '주행거리', '차량점검': '차량점검', '차량비고': '차량비고'
                }
                for i, col in enumerate(self.manual_visible_cols):
                    if col in translations:
                        self.manual_visible_cols[i] = translations[col]
                
                if self.manual_visible_cols and hasattr(self, 'daily_usage_tree'):
                    try:
                        self.daily_usage_tree['displaycolumns'] = [c for c in self.manual_visible_cols if c in self.daily_usage_tree['columns']]
                    except: pass
                
                # Restore monthly history visibility if saved
                self.monthly_visible_cols = config.get('monthly_visible_cols', [])
                if self.monthly_visible_cols and hasattr(self, 'monthly_usage_tree'):
                    try:
                        self.monthly_usage_tree['displaycolumns'] = [c for c in self.monthly_visible_cols if c in self.monthly_usage_tree['columns']]
                    except: pass

                # Restore budget site performance view visibility if saved
                self.budget_view_visible_cols = config.get('budget_view_visible_cols', [])
                self.budget_view_heading_aliases = config.get('budget_view_heading_aliases', {})
                self.budget_view_custom_columns = config.get('budget_view_custom_columns', [])
                if hasattr(self, 'budget_view_tree'):
                    try:
                        self._refresh_budget_view_tree_columns(reload_data=False)
                    except:
                        pass
                if self.budget_view_visible_cols and hasattr(self, 'budget_view_tree'):
                    try:
                        visible = [c for c in self.budget_view_visible_cols if c in self.budget_view_tree['columns']]
                        self.budget_view_tree['displaycolumns'] = visible if visible else self.budget_view_tree['columns']
                    except:
                        pass
                
                # Restore warehouses list
                self.warehouses[:] = config.get('warehouses', [])
                if not self.warehouses and not self.materials_df.empty:
                    self.warehouses = sorted(self.materials_df['창고'].dropna().unique().tolist())
                    self.warehouses = [str(w).strip() for w in self.warehouses if str(w).strip()]
                
                self.equipments[:] = config.get('equipments', [])
                if not self.equipments and not self.daily_usage_df.empty and '장비명' in self.daily_usage_df.columns:
                    self.equipments = sorted(self.daily_usage_df['장비명'].dropna().unique().tolist())
                    self.equipments = [str(e).strip() for e in self.equipments if str(e).strip()]
                
                # Vehicle list
                self.vehicles[:] = config.get('vehicles', [])
                
                # [FIX] Update combo box values after loading lists from config
                # Update sites combo boxes
                if hasattr(self, 'cb_trans_site'): 
                    self.cb_trans_site['values'] = sorted(self.sites)
                # Update warehouses combo boxes
                if hasattr(self, 'cb_warehouse'): 
                    self.cb_warehouse['values'] = sorted(self.warehouses)
                # Update users combo boxes
                if hasattr(self, 'ent_user'): 
                    self.ent_user['values'] = sorted(self.users)
                
                # Restore stock column widths
                stock_col_widths = config.get('stock_col_widths', {})
                if stock_col_widths and hasattr(self, 'stock_tree'):
                    for col, width in stock_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.stock_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass

                # Restore in/out history column widths
                inout_col_widths = config.get('inout_col_widths', {})
                if inout_col_widths and hasattr(self, 'inout_tree'):
                    for col, width in inout_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.inout_tree.column(col, width=w, minwidth=50, stretch=False)
                        except:
                            pass
                
                # Restore daily usage column widths
                daily_usage_col_widths = config.get('daily_usage_col_widths', {})
                if daily_usage_col_widths and hasattr(self, 'daily_usage_tree'):
                    for col, width in daily_usage_col_widths.items():
                        try:
                            # Only apply if width is reasonable (e.g. > 10)
                            w = int(width)
                            
                            # All columns are now user-resizable
                            if w > 10:
                                # Enforce minimums for high-precision cols
                                if col == '날짜': w = max(w, 160)
                                elif col == '입력시간': w = max(w, 300)
                                self.daily_usage_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass

                # Restore monthly usage column widths
                monthly_usage_col_widths = config.get('monthly_usage_col_widths', {})
                if monthly_usage_col_widths and hasattr(self, 'monthly_usage_tree'):
                    for col, width in monthly_usage_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.monthly_usage_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass
                
                # [NEW] Restore site summary column widths
                site_summary_col_widths = config.get('site_summary_col_widths', {})
                if site_summary_col_widths and hasattr(self, 'site_summary_tree'):
                    for col, width in site_summary_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.site_summary_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass

                # Restore budget site performance column widths
                budget_view_col_widths = config.get('budget_view_col_widths', {})
                if budget_view_col_widths and hasattr(self, 'budget_view_tree'):
                    for col, width in budget_view_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.budget_view_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass

                # [NEW] Restore worker summary column widths
                worker_summary_col_widths = config.get('worker_summary_col_widths', {})
                if worker_summary_col_widths and hasattr(self, 'worker_summary_tree'):
                    for col, width in worker_summary_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.worker_summary_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass
                # [NEW] Restore report column widths
                report_col_widths = config.get('report_col_widths', {})
                if report_col_widths and hasattr(self, 'report_tree'):
                    for col, width in report_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.report_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass
                
                if hasattr(self, '_loading_memos'):
                    del self._loading_memos
                    
                # Restore layout lock state
                if hasattr(self, 'btn_lock_layout'):
                    if self.layout_locked:
                        self.btn_lock_layout.config(text="🔒 배치 고정됨")
                        self.style.configure("Lock.TButton", foreground="black")
                    else:
                        self.btn_lock_layout.config(text="🔓 배치 수정 중")
                        self.style.configure("Lock.TButton", foreground="red")

                if config.get('daily_usage_sash_locked', False):
                    self.daily_usage_sash_locked = True
                    
                    if hasattr(self, 'btn_sash_lock'):
                        self.btn_sash_lock.config(text="🔒 경계 잠금됨")
                    
                    print("LOADED: Restored sash lock state")

                # Tab selection handled already at line 4531
                
                # Recreate Memos and Clones first (these must exist before they can be placed)
                self._loading_memos = []
                # Map class names to actual classes for recreation
                class_map = {'Entry': ttk.Entry, 'Combobox': ttk.Combobox}
                
                draggable_geos = config.get('draggable_geometries', {})
                for key, geo in draggable_geos.items():
                    if key.startswith('memo_'):
                        self._loading_memos.append(key)
                        self.add_new_memo(
                            initial_text=geo.get('text', ""), 
                            initial_title=geo.get('memo_title', "메모"), 
                            key=key
                        )
                    elif key.startswith('checklist_'):
                        self._loading_memos.append(key)
                        self.add_new_checklist(
                            initial_data=geo.get('checklist_items', []),
                            initial_title=geo.get('checklist_title', "체크리스트"),
                            key=key
                        )
                    elif key.startswith('clone_'):
                        self._loading_memos.append(key)
                        cls_name = geo.get('widget_class_name', 'Entry')
                        cls = class_map.get(cls_name, ttk.Entry)
                        label = geo.get('custom_label', "복제항목")
                        kwargs = geo.get('widget_kwargs', {})
                        m_list_key = geo.get('manage_list_key')
                        cont, w = self.create_draggable_container(self.entry_inner_frame, label, cls, key, manage_list_key=m_list_key, **kwargs)
                    elif key.startswith('vehicle_inspection_'):
                        self._loading_memos.append(key)
                        data = geo.get('vehicle_data', {})
                        self.add_vehicle_inspection_box(initial_data=data, key=key)

                # 2. DELAYED RESTORATION: Restore complex states after UI mapped
                def delayed_restore():
                    self.root.update_idletasks()
                    
                    # 1. Restore window state and size
                    try:
                        if config.get('window_state'):
                            if config['window_state'] == 'zoomed':
                                self.root.state('zoomed')
                            else:
                                try:
                                    self.root.state(config['window_state'])
                                except: pass
                        
                        w = config.get('window_width')
                        h = config.get('window_height')
                        if w and h:
                            # Only set geometry if not zoomed
                            if self.root.state() != 'zoomed':
                                self.root.geometry(f"{w}x{h}")
                    except: pass

                    # 2. Restore resolution lock if active
                    if self.resolution_locked:
                        self.root.resizable(False, False)
                        if hasattr(self, 'btn_resolution_lock'):
                            self.btn_resolution_lock.config(text="🔒 해상도 고정됨")

                    # 3. Restore draggable components regardless of lock state
                    # [FIX] Always restore positions if they exist, so movement is persisted 
                    # even if the app was closed while 'unlocked'.
                    draggable_geos = config.get('draggable_geometries', {})
                    for key, geo in draggable_geos.items():
                        if key in self.draggable_items:
                            widget = self.draggable_items[key]
                            try:
                                # [LAYOUT FIX] 핵심 박스는 저장 좌표를 무시하고 기본 grid 배치 사용
                                if key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
                                    continue

                                if geo.get('hidden'):
                                    # [STABILITY] If it's a core widget, ignore hidden status and reset it
                                    if key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
                                        print(f"LOADER: Recovering hidden core widget: {key}")
                                        self.reset_widget_position(None, widget=widget)
                                        continue
                                        
                                    if widget.winfo_manager() == 'grid':
                                        widget.grid_forget()
                                    elif widget.winfo_manager() == 'place':
                                        widget.place_forget()
                                    continue
                                
                                if geo.get('custom_label') and hasattr(widget, '_label_widget'):
                                    widget._label_widget.config(text=geo['custom_label'])
                                
                                # [FIX] Explicitly un-grid before using place manager
                                if widget.winfo_manager() == 'grid':
                                    if not hasattr(widget, '_original_grid_info'):
                                        widget._original_grid_info = widget.grid_info()
                                    widget.grid_forget()
                                
                                self._ensure_placeholder(widget, width=geo.get('width'), height=geo.get('height'))
                                widget.lift()
                                
                                # [FIX] Enforce minimum dimensions to prevent invisible widgets
                                load_x = geo.get('x', 0)
                                load_y = geo.get('y', 0)
                                load_w = max(100, geo.get('width', 200))
                                load_h = max(50, geo.get('height', 100))

                                # [RECOVERY] 작업자 박스는 화면 밖/과소 복원을 방지
                                if key == 'workers_box_geometry':
                                    try:
                                        p_w = widget.master.winfo_width()
                                        p_h = widget.master.winfo_height()
                                        load_w = max(520, load_w)
                                        load_h = max(360, load_h)
                                        if p_w > 0:
                                            load_x = max(0, min(int(load_x), max(0, p_w - load_w)))
                                        else:
                                            load_x = max(0, int(load_x))
                                        if p_h > 0:
                                            load_y = max(0, min(int(load_y), max(0, p_h - load_h)))
                                        else:
                                            load_y = max(0, int(load_y))
                                    except Exception:
                                        load_x = max(0, int(load_x))
                                        load_y = max(0, int(load_y))
                                
                                widget.place(x=load_x, y=load_y, width=load_w, height=load_h)

                                # [VISUAL FIX] 핵심 박스는 기존 placeholder 잔상 제거 + 최소 필요 높이 보정
                                if key in ('ndt_usage_box_geometry', 'rtk_usage_box_geometry', 'workers_box_geometry'):
                                    try:
                                        self._remove_placeholder(widget)
                                    except Exception:
                                        pass
                                    try:
                                        widget.update_idletasks()
                                        if key == 'workers_box_geometry':
                                            min_w = max(520, widget.winfo_reqwidth() + 8)
                                            min_h = max(360, widget.winfo_reqheight() + 12)
                                        else:
                                            min_w = max(160, widget.winfo_reqwidth() + 4)
                                            min_h = max(90, widget.winfo_reqheight() + 6)
                                        cur_x = int(float(widget.place_info().get('x', load_x)))
                                        cur_y = int(float(widget.place_info().get('y', load_y)))
                                        cur_w = int(float(widget.place_info().get('width', load_w)))
                                        cur_h = int(float(widget.place_info().get('height', load_h)))
                                        if cur_w < min_w or cur_h < min_h:
                                            widget.place(x=cur_x, y=cur_y, width=max(cur_w, min_w), height=max(cur_h, min_h))
                                        if key == 'workers_box_geometry':
                                            widget.lift()
                                    except Exception:
                                        pass
                            except Exception as e:
                                print(f"Error placing widget {key}: {e}")
                                
                    # 3.5 [LAYOUT FIX] 핵심 박스는 항상 기본 grid 위치로 복구
                    for key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
                        if key not in self.draggable_items:
                            continue
                        widget = self.draggable_items[key]
                        try:
                            self.reset_widget_position(None, widget=widget)
                            self._remove_placeholder(widget)
                            widget.lift()
                        except Exception as e:
                            print(f"LOADER: Failed to reset core widget {key}: {e}")

                    # 3.8 [SAFEGUARD] 저장 버튼 강제 복구 (화면 밖/숨김/겹침 방지)
                    if 'save_btn_geometry' in self.draggable_items:
                        sbox = self.draggable_items['save_btn_geometry']
                        try:
                            sbox.update_idletasks()
                            smgr = sbox.winfo_manager()
                            sw = sbox.winfo_width()
                            sh = sbox.winfo_height()

                            # Hidden/unmanaged or too small => reset to grid default first
                            if smgr == '' or sw < 80 or sh < 20:
                                self.reset_widget_position(None, widget=sbox)
                                smgr = sbox.winfo_manager()

                            # If placed, clamp into visible area and keep above siblings
                            if smgr == 'place':
                                pi = sbox.place_info()
                                sx = int(float(pi.get('x', 10)))
                                sy = int(float(pi.get('y', 340)))
                                sw = int(float(pi.get('width', max(120, sbox.winfo_reqwidth() or 120))))
                                sh = int(float(pi.get('height', max(28, sbox.winfo_reqheight() or 28))))
                                p_w = sbox.master.winfo_width()
                                p_h = sbox.master.winfo_height()
                                sw = max(120, sw)
                                sh = max(28, sh)
                                if p_w > 0:
                                    sx = max(0, min(sx, max(0, p_w - sw)))
                                if p_h > 0:
                                    sy = max(0, min(sy, max(0, p_h - sh)))
                                sbox.place(x=sx, y=sy, width=sw, height=sh)

                            sbox.lift()
                        except Exception:
                            pass

                    # 4. Restore sash positions
                    def apply_sashes():
                        try:
                            # Daily usage sash
                            daily_sash = config.get('daily_usage_sash_pos')
                            if daily_sash is not None and hasattr(self, 'daily_usage_paned'):
                                total_h = self.daily_usage_paned.winfo_height()
                                if total_h > 0:
                                    min_pos = max(int(total_h * 0.1), 460)
                                    max_pos = int(total_h * 0.9)
                                    safe_pos = int(daily_sash)
                                    if safe_pos < min_pos:
                                        safe_pos = min_pos
                                    if safe_pos > max_pos:
                                        safe_pos = max_pos
                                    self.daily_usage_paned.sashpos(0, safe_pos)
                                else:
                                    self.daily_usage_paned.sashpos(0, int(daily_sash))
                            
                            # History sash
                            history_sash = config.get('daily_history_sash_pos')
                            if history_sash is not None and hasattr(self, 'daily_history_paned'):
                                self.daily_history_paned.sashpos(0, int(history_sash))
                                
                            # If sash lock is active (check config heavily)
                            config_locked = config.get('daily_usage_sash_locked', False)
                            if config_locked:
                                self.daily_usage_sash_locked = True
                                if hasattr(self, 'btn_sash_lock'):
                                    self.btn_sash_lock.config(text="🔒 경계 고정됨")
                                    self.btn_sash_lock.configure(style="SashLock.TButton")
                                self._start_sash_monitor()
                            elif self.daily_usage_sash_locked: # If self says locked but config didn't (fallback)
                                self._start_sash_monitor()
                        except: pass

                    self.root.after(100, apply_sashes)
                    self.root.after(800, apply_sashes)

                    # 5. [FIX] Recalculate entry_inner_frame height instead of using potentially stale config
                    if hasattr(self, 'entry_inner_frame'):
                        self._adjust_parent_height(self.entry_inner_frame, force=True)
                        # Also refresh canvas
                        self._ensure_canvas_scroll_region()
                    
                    # 6. Final UI refresh
                    for l_key in ['users', 'sites', 'equipments', 'warehouses', 'worktimes']:
                        self.refresh_ui_for_list_change(l_key)

                    # 7. Mark as ready for future saves AFTER all restoration is done
                    def finalize_loading():
                        try:
                            print("[STARTUP] Finalizing loading...")
                            self.is_ready = True
                            
                            # [FIX] Forcibly show and deiconify window to prevent "closing downwards" or disappearing
                            print("[STARTUP] Ensuring window visibility...")
                            try:
                                self.root.deiconify()
                                self.root.lift()
                                self.root.focus_force()
                            except: pass
                            
                            # [V13_RESET_ON_STARTUP] Ensure Daily Usage form starts BLANK as requested
                            print("[STARTUP] Resetting forms...")
                            try:
                                self.clear_daily_usage_form_all()
                            except: pass
                            
                            # [NEW] Trigger column auto-hide on both tabs immediately after startup
                            print("[STARTUP] Refreshing views...")
                            try:
                                self.root.after(100, lambda: self.update_daily_usage_view())
                                self.root.after(200, lambda: self.update_monthly_usage_view())
                                self.root.after(300, lambda: self.update_budget_site_view()) 
                                # [NEW] Set default tab to Daily Usage (Index 4) on startup
                                self.root.after(400, lambda: self.notebook.select(4))
                            except Exception as e:
                                print(f"[STARTUP] View refresh error: {e}")
                                
                            print("APP READY: State restoration complete.")
                        except Exception as startup_err:
                            print(f"[CRITICAL STARTUP ERROR] {startup_err}")
                            messagebox.showerror("Startup Error", f"An error occurred during startup: {startup_err}\n\n{traceback.format_exc()}")
                    
                    print("[STARTUP] Scheduling finalization...")
                    self.root.after(500, finalize_loading) 
                
                # Execute delayed restoration
                try:
                    self.root.after(300, delayed_restore)
                except Exception as outer_err:
                    messagebox.showerror("Initialization Error", f"Critical error during initialization: {outer_err}")

        except Exception as e:
            print(f"Failed to load tab config: {e}")
            # Ensure we eventually become ready even if load failed, to allow new saves
            self.root.after(2000, lambda: setattr(self, 'is_ready', True))
    
    def on_closing(self):
        """Handle window closing event"""
        self.save_tab_config()
        self.root.destroy()
    
    def export_stock_to_excel(self):
        """Export current stock data to Excel"""
        try:
            # Get current filtered data from treeview
            stock_data = []
            
            for item in self.stock_tree.get_children():
                values = self.stock_tree.item(item, 'values')
                stock_data.append({
                    'ID': values[0],
                    '회사코드': values[1],
                    '관리품번': values[2],
                    '품목명': values[3],
                    'SN': values[4],
                    '창고': values[5],
                    '모델명': values[6],
                    '규격': values[7],
                    '품목군코드': values[8],
                    '공급업체': values[9],
                    '제조사': values[10],
                    '제조국': values[11],
                    '가격': values[12],
                    '관리단위': values[13],
                    '수량': values[14],
                    '재고하한': values[15]
                })
            
            if not stock_data:
                messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
                return
            
            # Prepare filename with current date
            current_date = datetime.datetime.now().strftime('%Y%m%d')
            filename = f"재고현황_{current_date}.xlsx"
            
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=filename,
                title="재고 현황 저장",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if save_path:
                stock_df = pd.DataFrame(stock_data)
                stock_df = self.clean_df_export(stock_df)
                self.save_df_to_excel_autofit(stock_df, save_path, "재고현황")
                messagebox.showinfo("완료", f"재고 현황이 저장되었습니다.\n저장 위치: {save_path}")
                
        except Exception as e:
            messagebox.showerror("오류", f"내보내기 실패: {e}")

    def enable_tree_column_drag(self, tree, context_menu_handler=None):
        """Ultra-robust drag-and-drop column reordering using root-level event tracking"""
        tree._drag_info = {"col_id": None, "col_name": None, "start_x": 0, "active": False, "ghost": None}
        
        def on_press(event):
            region = tree.identify_region(event.x, event.y)
            if region == "heading":
                col_id = tree.identify_column(event.x)
                col_name = self._get_column_name_from_id(tree, col_id)
                if col_name:
                    tree._drag_info.update({
                        "col_id": col_id,
                        "col_name": col_name,
                        "start_x": event.x_root,
                        "active": False,
                        "motion_id": self.root.bind("<B1-Motion>", on_motion, add="+"),
                        "release_id": self.root.bind("<ButtonRelease-1>", on_release, add="+")
                    })
                
        def on_motion(event):
            di = tree._drag_info
            if not di["col_name"]: return
            
            # Start drag if moved > 5 pixels
            if not di["active"] and abs(event.x_root - di["start_x"]) > 5:
                di["active"] = True
                tree.config(cursor="fleur")
                
                # Visual ghost label
                try:
                    if di["ghost"]: di["ghost"].destroy()
                    di["ghost"] = tk.Toplevel(self.root)
                    di["ghost"].overrideredirect(True)
                    di["ghost"].attributes("-alpha", 0.8)
                    di["ghost"].attributes("-topmost", True)
                    
                    lbl = tk.Label(di["ghost"], text=f" ↔ {di['col_name']} 이동 중 ", 
                                   bg="#424242", fg="white", relief="raised", borderwidth=1, 
                                   padx=8, pady=4, font=('Malgun Gothic', 10, 'bold'))
                    lbl.pack()
                except: pass
            
            if di["active"] and di["ghost"]:
                try:
                    # Follow mouse with offset to stay visible
                    di["ghost"].geometry(f"+{event.x_root + 20}+{event.y_root + 10}")
                except: pass

        def on_release(event):
            di = tree._drag_info
            if di["active"] and di["col_name"]:
                # Convert screen X back to tree-relative X
                tree_x = event.x_root - tree.winfo_rootx()
                target_id = tree.identify_column(tree_x)
                
                if target_id and target_id != di["col_id"]:
                    self._reorder_tree_columns(tree, di["col_id"], target_id)
            
            # Cleanup all bindings and state
            tree.config(cursor="")
            if di["ghost"]:
                try: di["ghost"].destroy()
                except: pass
                di["ghost"] = None
            
            # Unbind specific root-level drag listeners securely
            try:
                if di.get("motion_id"):
                    self.root.unbind("<B1-Motion>", di["motion_id"])
                if di.get("release_id"):
                    self.root.unbind("<ButtonRelease-1>", di["release_id"])
            except: pass
            
            di.update({"col_id": None, "col_name": None, "active": False, "motion_id": None, "release_id": None})

        # Primary entry point: tree-level press
        tree.bind("<Button-1>", on_press, add="+")
        # Direct right-click menu as reliable alternative
        if context_menu_handler is not None:
            tree.bind("<Button-3>", context_menu_handler, add="+")
        else:
            tree.bind("<Button-3>", lambda e: self._show_heading_context_menu(e, tree), add="+")

    def _show_heading_context_menu(self, event, tree):
        """Show context menu on Treeview header to move columns"""
        region = tree.identify_region(event.x, event.y)
        if region == "heading":
            column_id = tree.identify_column(event.x)
            
            # Identify the column name from visual ID
            col_name = self._get_column_name_from_id(tree, column_id)
            if not col_name: return

            menu = tk.Menu(self.root, tearoff=0)
            menu.add_command(label=f"⬅️ '{col_name}' 왼쪽으로 이동", 
                             command=lambda: self._move_column_visual(tree, column_id, -1))
            menu.add_command(label=f"➡️ '{col_name}' 오른쪽으로 이동", 
                             command=lambda: self._move_column_visual(tree, column_id, 1))
            menu.add_separator()
            menu.add_command(label="⚙️ 컬럼 관리(숨기기/보이기)...", command=self.show_column_visibility_dialog)
            
            menu.post(event.x_root, event.y_root)

    def _show_generic_tree_heading_context_menu(self, event, tree):
        """Generic header context menu for trees that support only column move."""
        try:
            if tree.identify_region(event.x, event.y) != "heading":
                return
            column_id = tree.identify_column(event.x)
            col_name = self._get_column_name_from_id(tree, column_id)
            if not col_name:
                return
        except Exception:
            return

        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label=f"⬅️ '{col_name}' 왼쪽으로 이동",
                         command=lambda: self._move_column_visual(tree, column_id, -1))
        menu.add_command(label=f"➡️ '{col_name}' 오른쪽으로 이동",
                         command=lambda: self._move_column_visual(tree, column_id, 1))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _get_column_name_from_id(self, tree, column_id):
        """Helper to get actual column name from visual ID like #1"""
        try:
            if not column_id or not column_id.startswith("#"): return None
            # Extract N from #N
            vis_idx = int(column_id[1:]) - 1
            
            # Get displaycolumns accurately
            disp_tuple = tree.cget("displaycolumns")
            if not disp_tuple or disp_tuple == ("#all",) or disp_tuple == ["#all"] or disp_tuple == "":
                visible_names = list(tree.cget("columns"))
            else:
                visible_names = list(disp_tuple)
            
            if 0 <= vis_idx < len(visible_names):
                return visible_names[vis_idx]
        except: pass
        return None

    def _move_column_visual(self, tree, source_id, direction):
        """Move a column left (-1) or right (1) based on its current position"""
        col_name = self._get_column_name_from_id(tree, source_id)
        if not col_name: return
        
        try:
            disp_tuple = tree.cget("displaycolumns")
            if not disp_tuple or disp_tuple == ("#all",) or disp_tuple == ["#all"] or disp_tuple == "":
                visible_names = list(tree.cget("columns"))
            else:
                visible_names = list(disp_tuple)

            if col_name not in visible_names: return
            
            src_idx = visible_names.index(col_name)
            tgt_idx = src_idx + direction
            
            if 0 <= tgt_idx < len(visible_names):
                visible_names.pop(src_idx)
                visible_names.insert(tgt_idx, col_name)
                tree["displaycolumns"] = visible_names
                
                # Update persistence
                if tree == self.daily_usage_tree:
                    self.manual_visible_cols = visible_names
                    self.save_tab_config()
                elif tree == self.monthly_usage_tree:
                    self.monthly_visible_cols = visible_names
                    self.save_tab_config()
                elif hasattr(self, 'budget_view_tree') and tree == self.budget_view_tree:
                    self.budget_view_visible_cols = visible_names
                    self.save_tab_config()
        except: pass

    def _reorder_tree_columns(self, tree, source_id, target_id):
        """Internal helper to swap displaycolumns for drag-and-drop"""
        src_name = self._get_column_name_from_id(tree, source_id)
        tgt_name = self._get_column_name_from_id(tree, target_id)
        
        if not src_name or not tgt_name or src_name == tgt_name: return
        
        try:
            disp_tuple = tree.cget("displaycolumns")
            if not disp_tuple or disp_tuple == ("#all",) or disp_tuple == ["#all"] or disp_tuple == "":
                visible_names = list(tree.cget("columns"))
            else:
                visible_names = list(disp_tuple)
            
            if src_name in visible_names and tgt_name in visible_names:
                src_idx = visible_names.index(src_name)
                tgt_idx = visible_names.index(tgt_name)
                
                # Strategic insertion: move before or after target 
                # (Simple swap approach: just pop and insert)
                visible_names.pop(src_idx)
                visible_names.insert(tgt_idx, src_name)
                
                tree["displaycolumns"] = visible_names
                
                # Update persistence
                if tree == self.daily_usage_tree:
                    self.manual_visible_cols = visible_names
                    self.save_tab_config()
                elif tree == self.monthly_usage_tree:
                    self.monthly_visible_cols = visible_names
                    self.save_tab_config()
                elif hasattr(self, 'budget_view_tree') and tree == self.budget_view_tree:
                    self.budget_view_visible_cols = visible_names
                    self.save_tab_config()
        except: pass


    def treeview_sort_column(self, tv, col, reverse):
        """Sort treeview contents when a column header is clicked"""
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        
        # Separate the 'Total' row (tagged 'total') from sorting
        data_rows = []
        total_row = None
        
        for val, k in l:
            # Check tags for this item
            tags = tv.item(k, 'tags')
            # If tags is a tuple/list, check if 'total' is in it. 
            # If it's a string (though usually tuple), check equality or contains.
            if tags and ('total' in tags or tags == 'total'):
                total_row = (val, k)
            else:
                data_rows.append((val, k))
                
        # Helper for numeric conversion
        def convert(val):
            try:
                # Remove common non-numeric chars
                s = str(val).replace(',', '').replace('시간', '').replace('Hrs', '').replace('원', '').replace(' ', '').replace('(', '').replace(')', '')
                if not s: return 0.0
                return float(s)
            except ValueError:
                return str(val).lower() # Default to string sort

        try:
            data_rows.sort(key=lambda t: convert(t[0]), reverse=reverse)
        except Exception:
            data_rows.sort(key=lambda t: t[0].lower(), reverse=reverse)

        # Rearrange items in sorted positions
        for index, (val, k) in enumerate(data_rows):
            tv.move(k, '', index)
            
        # Ensure Total row is always at the bottom
        if total_row:
             tv.move(total_row[1], '', 'end')

        # Reverse sort next time
        tv.heading(col, command=lambda: self.treeview_sort_column(tv, col, not reverse))

    def switch_to_daily_usage_details(self):
        """Switch to Daily Usage tab and apply current filters from Sales tab"""
        # Sync filters before switching

        site = self.cb_sales_filter_site.get().strip()
        start = self.sales_start_date.get_date()
        end = self.sales_end_date.get_date()
        
        # 1. Switch Tab (index 5)
        self.notebook.select(5)
        
        # [NEW] Force focus transition
        if hasattr(self, 'daily_usage_tree'):
            self.daily_usage_tree.focus_set()

        # 2. Sync Filters
        if hasattr(self, 'cb_daily_filter_site'):
            self.cb_daily_filter_site.set(site)
        
        if hasattr(self, 'ent_daily_start_date'):
            self.ent_daily_start_date.set_date(start)
            
        if hasattr(self, 'ent_daily_end_date'):
            self.ent_daily_end_date.set_date(end)
            
        # 3. Trigger Search
        self.update_daily_usage_view()

    def save_all_daily_usage_changes(self):
        """Unified save button for the inquiry list to satisfy spreadsheet-style workflow"""
        try:
            if self.save_data():
                # Provide clear feedback that ALL changes in the app are saved
                messagebox.showinfo("저장 완료", "모든 변경 사항이 마스터 파일(Material_Inventory.xlsx)에 안전하게 저장되었습니다.")
                # Refresh view to ensure everything is synchronized
                self.update_daily_usage_view()
                self.refresh_inquiry_filters()
        except Exception as e:
            messagebox.showerror("저장 오류", f"데이터 저장 중 오류가 발생했습니다: {e}")

    def open_detached_monthly_usage_view(self):
        """월별 집계 조회를 별도의 팝업창(모니터링 창)으로 엽니다. (메인 화면과 100% 동일한 구성)"""
        # 기존 창이 있으면 포커스만 이동
        if 'monthly' in self.detached_windows and self.detached_windows['monthly']['window'].winfo_exists():
            self.detached_windows['monthly']['window'].lift()
            return

        popup = tk.Toplevel(self.root)
        popup.title("📊 월별 집계 모니터링 (팝업)")
        popup.geometry("1600x900")
        
        main_frame = ttk.Frame(popup, padding=10)
        main_frame.pack(expand=True, fill='both')
        
        # 상단 정보/필터 안내
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill='x', pady=(0, 5))
        
        # [NEW] Popup Search Filters
        filter_row = ttk.Frame(info_frame)
        filter_row.pack(side='left', fill='x', expand=True)

        def sync_and_refresh(event=None):
            # Mirror popup filters to main window
            self.cb_filter_year.set(p_cb_year.get())
            self.cb_filter_month.set(p_cb_month.get())
            self.cb_filter_site_monthly.set(p_cb_site.get())
            self.cb_filter_material_monthly.set(p_cb_mat.get())
            self.update_monthly_usage_view()

        ttk.Label(filter_row, text="연도:").pack(side='left', padx=2)
        p_cb_year = ttk.Combobox(filter_row, values=self.cb_filter_year['values'], width=8)
        p_cb_year.pack(side='left', padx=2)
        p_cb_year.set(self.cb_filter_year.get())
        p_cb_year.bind("<<ComboboxSelected>>", sync_and_refresh)

        ttk.Label(filter_row, text="월:").pack(side='left', padx=2)
        p_cb_month = ttk.Combobox(filter_row, values=self.cb_filter_month['values'], width=6)
        p_cb_month.pack(side='left', padx=2)
        p_cb_month.set(self.cb_filter_month.get())
        p_cb_month.bind("<<ComboboxSelected>>", sync_and_refresh)

        ttk.Label(filter_row, text="현장:").pack(side='left', padx=2)
        p_cb_site = ttk.Combobox(filter_row, values=self.cb_filter_site_monthly['values'], width=12)
        p_cb_site.pack(side='left', padx=2)
        p_cb_site.set(self.cb_filter_site_monthly.get())
        p_cb_site.bind("<<ComboboxSelected>>", sync_and_refresh)

        ttk.Label(filter_row, text="품목:").pack(side='left', padx=2)
        p_cb_mat = ttk.Combobox(filter_row, values=self.cb_filter_material_monthly['values'], width=18)
        p_cb_mat.pack(side='left', padx=2)
        p_cb_mat.set(self.cb_filter_material_monthly.get())
        p_cb_mat.bind("<<ComboboxSelected>>", sync_and_refresh)

        ttk.Button(filter_row, text="조회", width=6, command=sync_and_refresh).pack(side='left', padx=5)
        
        paned = ttk.PanedWindow(main_frame, orient="vertical")
        paned.pack(expand=True, fill='both')
        
        # 1. 메인 트리뷰 (상단)
        tree_frame = ttk.Frame(paned)
        paned.add(tree_frame, weight=3) 
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        columns = ('연도', '월', '현장', '작업자', '작업시간', 'OT시간', 'OT금액', 'OT1', 'OT2', 'OT3', 'OT4', 'OT5', 'OT6', 'OT7', 'OT8', 'OT9', 'OT10', 
                   '수량', '단가', '출장비', '일식', '검사비', '품목명', '센터미스', '농도', '마킹미스', '필름마크', 
                   '취급부주의', '고객불만', '기타', 'RTK총계', '형광자분', '흑색자분', '백색페인트', '침투제', '세척제', '현상제', '형광침투제', '비고', '(Full작업자)')
        
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree['displaycolumns'] = [c for c in columns if c != '(Full작업자)']
        
        vsb.config(command=tree.yview); hsb.config(command=tree.xview)
        
        col_widths = {'연도': 90, '월': 70, '현장': 140, '작업자': 100, '작업시간': 100, 'OT시간': 100, 'OT금액': 110, '품목명': 200, '비고': 220}
        for col in columns:
            tree.heading(col, text=col, command=lambda c=col: self.treeview_sort_column(tree, c, False))
            tree.column(col, width=col_widths.get(col, 100), anchor='center', stretch=False)
        
        tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1); tree_frame.grid_columnconfigure(0, weight=1)

        # [NEW] 인터랙티브 기능 복제
        self.enable_tree_column_drag(tree)
        tree.tag_configure('total', background='#E8F4F8', font=('Arial', 12, 'bold'))
        tree.bind("<Button-1>", lambda e: self.show_worker_popup(e, tree), add="+")
        tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())

        # 2. 현장별 요약 (중단 - 메인 UI는 3단 수직 분할)
        site_frame = ttk.LabelFrame(paned, text="현장별 누계")
        paned.add(site_frame, weight=1)
        site_cols = ('현장', '검사방법', '품목명', '수량', '검사비', '출장비', '형광자분', '흑색자분', '백색페인트', 
                     '침투제', '세척제', '현상제', '형광침투제', '센터미스', '농도', '마킹미스', '필름마크', '취급부주의', '고객불만', '기타', 'RTK총계')
        site_tree = ttk.Treeview(site_frame, columns=site_cols, show='headings')
        site_vsb = ttk.Scrollbar(site_frame, orient="vertical", command=site_tree.yview)
        site_tree.configure(yscrollcommand=site_vsb.set)
        for col in site_cols:
            site_tree.heading(col, text=col, command=lambda c=col: self.treeview_sort_column(site_tree, c, False))
            w = 120 if col in ['현장', '검사방법', '품목명'] else (100 if col in ['검사비', '출장비'] else 80)
            site_tree.column(col, width=w, anchor='center', stretch=False)
        site_tree.pack(side='left', expand=True, fill='both')
        site_vsb.pack(side='right', fill='y')
        
        self.enable_tree_column_drag(site_tree, context_menu_handler=lambda e: self._show_generic_tree_heading_context_menu(e, site_tree))
        site_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())
        
        # 3. 작업자별 요약 (하단)
        worker_frame = ttk.LabelFrame(paned, text="작업자별 누계")
        paned.add(worker_frame, weight=1)
        worker_cols = ('작업자', '총공수', '연장(시간)', '야간(시간)', '휴일(시간)', '총OT(시간)', '연장(금액)', '야간(금액)', '휴일(금액)', '총OT(금액)')
        worker_tree = ttk.Treeview(worker_frame, columns=worker_cols, show='headings')
        worker_vsb = ttk.Scrollbar(worker_frame, orient="vertical", command=worker_tree.yview)
        worker_tree.configure(yscrollcommand=worker_vsb.set)
        worker_widths = {'작업자': 100, '총공수': 70, '총OT(시간)': 80, '총OT(금액)': 100}
        for col in worker_cols:
            worker_tree.heading(col, text=col, command=lambda c=col: self.treeview_sort_column(worker_tree, c, False))
            worker_tree.column(col, width=worker_widths.get(col, 80), anchor='center', stretch=False)
        worker_tree.pack(side='left', expand=True, fill='both')
        worker_vsb.pack(side='right', fill='y')

        self.enable_tree_column_drag(worker_tree, context_menu_handler=lambda e: self._show_generic_tree_heading_context_menu(e, worker_tree))
        worker_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())

        # [NEW] 메인 UI의 분할 배율 동기화
        def sync_sash():
            try:
                main_h = self.monthly_paned.winfo_height()
                if main_h > 0:
                    popup_h = paned.winfo_height()
                    if popup_h > 0:
                        s1 = self.monthly_paned.sashpos(0)
                        s2 = self.monthly_paned.sashpos(1)
                        paned.sashpos(0, int(s1 * popup_h / main_h))
                        paned.sashpos(1, int(s2 * popup_h / main_h))
            except: pass
        popup.after(500, sync_sash) # UI 렌더링 후 적용

        # 상태 등록
        self.detached_windows['monthly'] = {
            'window': popup,
            'tree': tree,
            'site_tree': site_tree,
            'worker_tree': worker_tree,
            'filters': {
                'year': p_cb_year,
                'month': p_cb_month,
                'site': p_cb_site,
                'mat': p_cb_mat
            }
        }
        
        def on_close():
            if 'monthly' in self.detached_windows:
                del self.detached_windows['monthly']
            popup.destroy()
        popup.protocol("WM_DELETE_WINDOW", on_close)
        
        # 데이터 복사
        self.update_monthly_usage_view()
        
        # [FIXED] Extraction using standardized indices (Year=0, Month=1, Site=2)
        def on_popup_select(event):
            selection = tree.selection()
            if not selection: return
            item = selection[0]
            try:
                vals = tree.item(item, 'values')
                if not vals: return
                tags = tree.item(item, 'tags')
                
                if not hasattr(self, 'current_monthly_df'): return
                
                if 'total' in tags:
                    subset = self.current_monthly_df
                else:
                    # Robust extraction
                    try:
                        y = int(vals[0]) 
                        m = int(vals[1])
                        s = str(vals[2]).strip()
                        
                        # Material name is at index 22 in our formatted tuple
                        mat_name = str(vals[22]).strip() if len(vals) > 22 else None
                        
                        mask = (self.current_monthly_df['Year'] == y) & \
                               (self.current_monthly_df['Month'] == m) & \
                               (self.current_monthly_df['Site'] == s)
                        
                        if mat_name:
                            ids = self.materials_df[self.materials_df['품목명'] == mat_name]['MaterialID'].tolist()
                            if ids:
                                mask = mask & (self.current_monthly_df['MaterialID'].isin(ids))
                        
                        subset = self.current_monthly_df[mask]
                    except:
                        subset = pd.DataFrame()
                
                # Update summaries
                self._populate_monthly_summary_trees(subset)
            except Exception as e:
                print(f"DEBUG: Detached selection error: {e}")

        tree.bind("<<TreeviewSelect>>", on_popup_select)
        
        if hasattr(self, 'current_monthly_df'):
            self._populate_monthly_summary_trees(self.current_monthly_df)

    def open_detached_budget_view(self):
        """공사실행예산서 탭을 별도의 팝업창으로 엽니다. (메인 화면과 유사한 KPI/목록 구성)"""
        if 'budget' in self.detached_windows and self.detached_windows['budget']['window'].winfo_exists():
            self.detached_windows['budget']['window'].lift()
            return

        popup = tk.Toplevel(self.root)
        popup.title("📋 공사 실행예산서 모니터링 (팝업)")
        popup.geometry("1400x800")
        
        main_frame = ttk.Frame(popup, padding=10)
        main_frame.pack(expand=True, fill='both')
        
        # 1. KPI Panel (메인 UI 그대로 재현)
        kpi_frame = tk.Frame(main_frame, bg='#f8fafc', height=80)
        kpi_frame.pack(fill='x', pady=(0, 10))
        kpi_frame.pack_propagate(False)
        
        def create_kpi_card(parent, title, color):
            frame = tk.Frame(parent, bg='white', highlightbackground='#e2e8f0', highlightthickness=1)
            frame.pack(side='left', expand=True, fill='both', padx=5, pady=5)
            tk.Label(frame, text=title, font=('Malgun Gothic', 9), fg='#64748b', bg='white').pack(pady=(5, 2))
            lbl = tk.Label(frame, text="0", font=('Malgun Gothic', 16, 'bold'), fg=color, bg='white')
            lbl.pack(pady=(0, 5))
            return lbl

        lbl_rev = create_kpi_card(kpi_frame, "총 매출액", "#2563eb")
        lbl_cost = create_kpi_card(kpi_frame, "총 실행원가", "#dc2626")
        lbl_profit = create_kpi_card(kpi_frame, "총 기대이익", "#059669")
        lbl_margin = create_kpi_card(kpi_frame, "평균 이익률", "#7c3aed")
        
        # 2. 예산 목록 트리뷰
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(expand=True, fill='both')
        
        vsb = ttk.Scrollbar(tree_frame, orient='vertical')
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal')
        
        columns = ('Site', 'Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit', 'Margin', 'Note')
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        col_map = {
            'Site': ('현장명', 150), 'Revenue': ('매출액', 120), 'UnitPrice': ('단가', 100),
            'LaborCost': ('노무비', 120), 'MaterialCost': ('자재비', 120), 'Expense': ('경비', 120),
            'OutsourceCost': ('외주비', 120), 'Profit': ('기대이익', 120), 'Margin': ('이익률', 100), 'Note': ('비고', 200)
        }
        for col in columns:
            head, width = col_map[col]
            tree.heading(col, text=head, command=lambda c=col: self.treeview_sort_column(tree, c, False))
            tree.column(col, width=width, anchor='center', stretch=False)
            
        tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1); tree_frame.grid_columnconfigure(0, weight=1)
        
        vsb.config(command=tree.yview); hsb.config(command=tree.xview)
        
        # [NEW] 인터랙티브 기능 추가
        self.enable_tree_column_drag(tree)

        self.detached_windows['budget'] = {
            'window': popup,
            'tree': tree,
            'lbl_rev': lbl_rev,
            'lbl_cost': lbl_cost,
            'lbl_profit': lbl_profit,
            'lbl_margin': lbl_margin
        }
        
        def on_close():
            if 'budget' in self.detached_windows:
                del self.detached_windows['budget']
            popup.destroy()
        popup.protocol("WM_DELETE_WINDOW", on_close)
        
        self._refresh_detached_budget()

    def _refresh_detached_budget(self):
        """팝업창의 예산 리스트와 KPI를 최신 데이터로 갱신"""
        if 'budget' not in self.detached_windows: return
        detached = self.detached_windows['budget']
        tree = detached['tree']
        
        for item in tree.get_children(): tree.delete(item)
        
        if self.budget_df.empty:
            for lbl in [detached['lbl_rev'], detached['lbl_cost'], detached['lbl_profit']]: lbl.config(text="0")
            detached['lbl_margin'].config(text="0.0%")
            return
        
        total_rev = total_cost = total_profit = 0
        
        for _, row in self.budget_df.iterrows():
            rev = pd.to_numeric(row.get('Revenue', 0), errors='coerce') or 0
            unit_p = pd.to_numeric(row.get('UnitPrice', 0), errors='coerce') or 0
            lab = pd.to_numeric(row.get('LaborCost', 0), errors='coerce') or 0
            mat = pd.to_numeric(row.get('MaterialCost', 0), errors='coerce') or 0
            exp = pd.to_numeric(row.get('Expense', 0), errors='coerce') or 0
            out = pd.to_numeric(row.get('OutsourceCost', 0), errors='coerce') or 0
            prof = pd.to_numeric(row.get('Profit', 0), errors='coerce') or 0
            
            total_rev += rev
            total_cost += (lab + mat + exp + out)
            total_profit += prof
            
            margin = (prof / rev * 100) if rev > 0 else 0
            
            tree.insert('', 'end', values=(
                row.get('Site', ''),
                f"{rev:,.0f}",
                f"{unit_p:,.0f}",
                f"{lab:,.0f}",
                f"{mat:,.0f}",
                f"{exp:,.0f}",
                f"{out:,.0f}",
                f"{prof:,.0f}",
                f"{margin:.1f}%",
                row.get('Note', '')
            ))
            
        # KPI Update
        detached['lbl_rev'].config(text=f"{total_rev:,.0f}")
        detached['lbl_cost'].config(text=f"{total_cost:,.0f}")
        detached['lbl_profit'].config(text=f"{total_profit:,.0f}")
        total_margin = (total_profit / total_rev * 100) if total_rev > 0 else 0
        detached['lbl_margin'].config(text=f"{total_margin:.1f}%")


    def cleanup_auto_transactions(self):
        """Remove existing 'Automatic Deduction' transactions for non-consumable equipment (PAUT, Yokes, etc.)"""
        if self.transactions_df.empty: 
            messagebox.showinfo("알림", "처리할 입출고 기록이 없습니다.")
            return
        
        confirm = messagebox.askyesno("클린업 확인", 
            "현재 저장된 '자동 차감' 기록 중 장비류(PAUT, Yoke 등)에 해당하는 내역만 골라 삭제하시겠습니까?\n\n"
            "※ 필름, 약품 등 소모품 기록은 유지됩니다.")
        if not confirm: return

        initial_count = len(self.transactions_df)
        
        def is_bad_auto(row):
            note = str(row.get('Note', ''))
            # [FIX] Match both English and potentially broken encoding versions of "(자동 차감)"
            # (ڵ ) is a common corruption of (자동 차감) when read as different encoding
            is_auto = "(자동 차감)" in note or "(ڵ )" in note or "(\ucac0 \u0020)" in note or "(\u00e1\u00b6\u00bd\u0020)" in note

            if not is_auto:
                return False
            
            mat_id = row.get('MaterialID')
            mat_name = self.get_material_display_name(mat_id)
            
            # If it's NOT a consumable, it's a candidate for removal
            # _is_consumable_material handles empty method by returning False for equipment keywords
            return not self._is_consumable_material(mat_name, "")

        mask = self.transactions_df.apply(is_bad_auto, axis=1)
        self.transactions_df = self.transactions_df[~mask]
        
        removed_count = initial_count - len(self.transactions_df)
        if removed_count > 0:
            if self.save_data():
                messagebox.showinfo("클린업 완료", f"장비류(PAUT, Yoke 등)의 자동 차감 내역 {removed_count}건을 삭제했습니다.")
                self.update_stock_view()
                self.update_transaction_view()
        else:
            messagebox.showinfo("알림", "정리할 대상이 없습니다.")

if __name__ == "__main__":
    root = tk.Tk()
    app = MaterialManager(root)
    try:
        root.mainloop()
    except KeyboardInterrupt:
        try:
            root.destroy()
        except Exception:
            pass
