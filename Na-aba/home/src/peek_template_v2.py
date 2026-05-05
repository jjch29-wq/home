from openpyxl import load_workbook
import os

path = os.path.abspath("../resources/Template_DailyWorkReport.xlsx")
wb = load_workbook(path)
sheet = wb.active

for r in range(1, 60):
    val = sheet[f'B{r}'].value
    if val:
        print(f"Row {r}: {val}")

# Also check merged cells in that range
print("\nMerged Cells in B18:S25 range:")
for rng in sheet.merged_cells.ranges:
    if rng.min_row >= 18 and rng.max_row <= 25:
        print(rng)

wb.close()
