from openpyxl import load_workbook
import os

path = os.path.abspath("../resources/Template_DailyWorkReport.xlsx")
wb = load_workbook(path)
sheet = wb.active

print("Peeking at rows 25-35:")
for r in range(25, 36):
    val = sheet[f'B{r}'].value
    if val:
        print(f"Row {r}: {val}")
    else:
        # Check if it's part of a merge
        for rng in sheet.merged_cells.ranges:
            if r >= rng.min_row and r <= rng.max_row and 2 >= rng.min_col and 2 <= rng.max_col:
                print(f"Row {r}: (Merged in {rng})")
                break

wb.close()
