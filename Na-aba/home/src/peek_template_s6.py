from openpyxl import load_workbook
import os

path = os.path.abspath("../resources/Template_DailyWorkReport.xlsx")
wb = load_workbook(path)
sheet = wb.active

print("Searching for Section 6 title:")
for r in range(40, 60):
    row_text = "".join([str(sheet.cell(row=r, column=c).value or '') for c in range(1, 5)])
    if "6." in row_text:
        print(f"Row {r}: {row_text}")

# Check merges in row 45 onwards
print("\nMerges around row 45-55:")
for rng in sheet.merged_cells.ranges:
    if rng.min_row >= 43 and rng.max_row <= 55:
        print(rng)

wb.close()
