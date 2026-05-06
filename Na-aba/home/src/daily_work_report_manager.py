import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.cell.cell import MergedCell, Cell
import os
import datetime

class DailyWorkReportManager:
    def __init__(self, template_path):
        self.template_path = template_path
        # Default mapping for fallback
        self.default_mapping = {
            'header': {'date': 'F2'},
            'general': {
                'company': 'E5', 'project_name': 'E6', 'standard': 'E7', 'equipment': 'E8',
                'report_no': 'K5', 'inspection_item': 'K6', 'inspector': 'K7', 'inspector_n8': 'N8', 'car_no': 'N9'
            },
            'methods': {
                'RT': {'row': 13}, 'UT': {'row': 14}, 'MT': {'row': 15}, 'PT': {'row': 16},
                'HT': {'row': 17}, 'VT': {'row': 18}, 'LT': {'row': 19}, 'ET': {'row': 20},
                'PAUT': {'row': 21}
            },
            'rtk': {
                'center_miss': 'D34', 'density': 'F34', 'marking_miss': 'H34', 'film_mark': 'J34',
                'handling': 'L34', 'customer_complaint': 'N34', 'etc': 'P34', 'total': 'R34'
            },
            'ot': {
                'row1_name': 'B38', 'row1_company': 'F38', 'row1_method': 'I38', 'row1_hours': 'K38', 'row1_amount': 'N38',
                'row2_name': 'B39', 'row2_company': 'F39', 'row2_method': 'I39', 'row2_hours': 'K39', 'row2_amount': 'N39'
            },
            'slim_widths': {
                'A': 1, 'B': 5, 'C': 1, 'D': 6, 'E': 2, 'F': 8, 'G': 1, 'H': 5, 
                'I': 4, 'J': 4, 'K': 5, 'L': 4, 'M': 5, 'N': 4, 'O': 5, 
                'P': 4, 'Q': 5, 'R': 4, 'S': 7
            },
            'materials': {
                'RT T200': 43, 'RT AA400': 44, 'RT Other': 45,
                'MT WHITE': 46, 'MT 7C-BLACK': 47,
                'PT Penetrant': 48, 'PT Cleaner': 49, 'PT Developer': 50
            },
            'vehicles': {
                'row_start': 26, 
                'col_map': {'vehicle_info': 'B', 'mileage': 'H', 'fuel': 'C', 'clean': 'E', 'oil': 'G', 'tire': 'I', 'light': 'K', 'safety': 'M'}
            }
        }

    def generate_report(self, data, output_path, custom_mapping=None):
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import re as _re
        from copy import copy
        import openpyxl.utils

        custom_mapping = custom_mapping or {}
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template not found at {self.template_path}")

        wb = openpyxl.load_workbook(self.template_path)
        sheet = wb.active
        mapping = custom_mapping if custom_mapping else self.default_mapping

        from openpyxl.cell.cell import Cell as RealCell
        from openpyxl.cell.read_only import EmptyCell
        from openpyxl.styles import Font, Alignment, Border, Side
        
        black_font = Font(name='맑은 고딕', size=10, bold=False, color='000000')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
        thin = Side(style='thin')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Helper 1: free a cell by unmerging (used BEFORE merge_cells)
        def free_cell(sheet, row, col):
            for m_range in list(sheet.merged_cells.ranges):
                if row >= m_range.min_row and row <= m_range.max_row and col >= m_range.min_col and col <= m_range.max_col:
                    sheet.unmerge_cells(m_range.coord)
            key = (row, col)
            if key not in sheet._cells or isinstance(sheet._cells[key], MergedCell):
                sheet._cells[key] = Cell(sheet, row=row, column=col)

        # Helper 2: write to cell via _cells dict WITHOUT unmerging (used AFTER merge_cells)
        def write_cell(sheet, row, col, value, font=None, align=None, border=None):
            key = (row, col)
            if key not in sheet._cells or isinstance(sheet._cells[key], MergedCell):
                sheet._cells[key] = Cell(sheet, row=row, column=col)
            cell = sheet._cells[key]
            cell.value = value
            if border: cell.border = border
            elif not cell.border or cell.border == Border(): cell.border = thin_border
            if font: cell.font = font
            if align: cell.alignment = align

        def safe_write(cell_coord, value, is_currency=False, is_bold=False):
            if not cell_coord: return
            try:
                cell = sheet[cell_coord]
                if isinstance(cell, MergedCell):
                    for m_range in sheet.merged_cells.ranges:
                        if cell_coord in m_range:
                            cell = sheet.cell(row=m_range.min_row, column=m_range.min_col)
                            break
                cell.value = value
                cell.font = Font(name='맑은 고딕', size=9, bold=is_bold)
                # [FIX] shrinkToFit and wrapText are mutually exclusive in Excel - use wrapText only
                cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                if is_currency:
                    cell.number_format = '#,##0 "원"'
            except: pass

        # 1. Header & General
        d = data.get('date', datetime.date.today())
        weekdays = ["월", "화", "수", "목", "금", "토", "일"]
        date_str = f"{d.year}년 {d.month}월 {d.day}일 ({weekdays[d.weekday()]})"
        safe_write(mapping.get('header', {}).get('date', 'F2'), date_str)

        gen_map = mapping.get('general', {})
        for key in ['company', 'project_name', 'standard', 'equipment', 'report_no', 'inspection_item', 'inspector', 'car_no', 'inspector_n8']:
            val_key = 'inspector' if key == 'inspector_n8' else key
            safe_write(gen_map.get(key), data.get(val_key, ''))

        # 2. Methods (Dynamic Offset)
        methods_data = data.get('methods', {})
        current_methods = list(methods_data.keys())
        method_offset = max(0, len(current_methods) - 4)
        
        import re as _re
        def fix_c(coord, offset):
            if not coord or not isinstance(coord, str): return coord
            match = _re.match(r"([A-Z]+)([0-9]+)", coord)
            if match:
                col, row = match.groups()
                row_val = int(row)
                if row_val >= 18: return f"{col}{row_val + offset}"
            return coord

        method_base_rows = [13, 14, 15, 16]
        if method_offset > 0:
            sheet.insert_rows(17, method_offset)
            for i in range(method_offset):
                target_r = 18 + method_offset + i
                if target_r <= 25 + method_offset:
                    sheet.row_dimensions[target_r].height = 5
            method_base_rows += list(range(17, 17 + method_offset))

        # Clear default method rows (13-16) first to ensure only saved data is shown
        for r_clear in [13, 14, 15, 16]:
            for c_clear in ['B', 'E', 'H', 'K', 'N', 'Q']:
                safe_write(f"{c_clear}{r_clear}", "")

        for idx, m_name in enumerate(current_methods):
            if idx < len(method_base_rows):
                r = method_base_rows[idx]; m = methods_data[m_name]
                safe_write(f'B{r}', m_name); safe_write(f'E{r}', m.get('unit', ''))
                safe_write(f'H{r}', m.get('qty', 0)); safe_write(f'K{r}', m.get('price', 0), is_currency=True)
                safe_write(f'N{r}', m.get('travel', 0), is_currency=True); safe_write(f'Q{r}', m.get('total', 0), is_currency=True)

        # 3. Section 2 (Notes) - Styling & Normalize
        h10 = sheet.row_dimensions[10].height if sheet.row_dimensions[10].height else 15
        sheet.row_dimensions[17].height = h10
        # 4. Section 4 (RTK Quality Defect Rate)
        rtk_header_row = None
        for r_search in range(25, 45): # Target common area
            for c_search in range(1, 20):
                cell_val = str(sheet.cell(row=r_search, column=c_search).value or "")
                if any(k in cell_val for k in ["품질", "RTK"]):
                    rtk_header_row = r_search
                    print(f"DEBUG: RTK Header found at row {rtk_header_row} (Col {c_search}, Val: '{cell_val}')")
                    break
            if rtk_header_row: break
        
        if rtk_header_row:
            rtk_label_row = rtk_header_row + 1
            found_label = False
            for r_chk in range(rtk_header_row, rtk_header_row + 5):
                for c_chk in range(1, 10):
                    if "센터미스" in str(sheet.cell(row=r_chk, column=c_chk).value or ""):
                        rtk_label_row = r_chk
                        found_label = True; break
                if found_label: break
            
            target_r = rtk_label_row + 1
            rtk_data = data.get('rtk', {})
            rtk_cols = {'센터미스': 4, '농도': 6, '마킹미스': 8, '필름마크': 10, '취급부주의': 12, '고객불만': 14, '기타': 16}
            for kor, col_idx in rtk_cols.items():
                val = rtk_data.get(kor, 0)
                write_cell(sheet, target_r, col_idx, val, black_font, center_align)
            
            rtk_total = data.get('rtk_total', rtk_data.get('총계', 0))
            display_total = f"{rtk_total}매" if rtk_total else "0매"
            write_cell(sheet, target_r, 18, display_total, black_font, center_align)
            write_cell(sheet, target_r, 19, display_total, black_font, center_align)
        
        # 5. Section 5 (OT)
        ot_header_row = None
        for r_search in range(1, 100):
            if sheet.cell(row=r_search, column=2).value == "검사자":
                ot_header_row = r_search; break
        
        if ot_header_row is None:
            ot_header_row = 37 + method_offset
        else:
            if ot_header_row != 37 + method_offset:
                ot_header_row = 37 + method_offset

        ot_list = data.get('ot_status', [])
        grouped_ot = {}
        for ot in ot_list:
            comp = str(ot.get('company', '')).strip()
            meth = str(ot.get('method', '')).strip()
            raw_h = str(ot.get('ot_hours', '')).replace('₩', '').replace(',', '').strip()
            clean_h = _re.sub(r'\(.*?\)', '', raw_h).strip()
            amt = str(ot.get('ot_amount', '')).strip()
            key = (comp, meth, clean_h, amt)
            if key not in grouped_ot:
                grouped_ot[key] = {'names': [], 'company': comp, 'method': meth, 'ot_hours': raw_h, 'ot_amount': amt}
            name = str(ot.get('names', '')).strip()
            if name and name not in grouped_ot[key]['names']:
                grouped_ot[key]['names'].append(name)
        
        final_ot_list = []
        for key in grouped_ot:
            item = grouped_ot[key]
            item['names_display'] = ", ".join(item['names'])
            final_ot_list.append(item)

        ot_count = len(final_ot_list)
        ot_extra = max(0, ot_count - 2)
        if ot_extra > 0:
            print(f"DEBUG: Inserting {ot_extra} OT rows and compressing Remarks section")
            # 1. Insert rows after 39
            for i in range(ot_extra):
                new_row_idx = 40 + method_offset + i
                source_row_idx = 39 + method_offset
                sheet.insert_rows(new_row_idx)
                
                # Copy style (Font, Alignment, Border, Fill) from source row
                import copy as _copy
                for c_idx in range(2, 20): # B to S
                    source_cell = sheet.cell(row=source_row_idx, column=c_idx)
                    target_cell = sheet.cell(row=new_row_idx, column=c_idx)
                    if source_cell.has_style:
                        target_cell.font = _copy.copy(source_cell.font)
                        target_cell.alignment = _copy.copy(source_cell.alignment)
                        target_cell.border = _copy.copy(source_cell.border)
                        target_cell.fill = _copy.copy(source_cell.fill)
                
                # Copy row height too
                sheet.row_dimensions[new_row_idx].height = sheet.row_dimensions[source_row_idx].height
                
                # [DYNAMIC MERGE COPY] Re-apply all merges from the source row to the new row
                for m_range in list(sheet.merged_cells.ranges):
                    if m_range.min_row == source_row_idx and m_range.max_row == source_row_idx:
                        new_range_coord = f"{openpyxl.utils.get_column_letter(m_range.min_col)}{new_row_idx}:{openpyxl.utils.get_column_letter(m_range.max_col)}{new_row_idx}"
                        try: sheet.merge_cells(new_range_coord)
                        except: pass

            # 2. Compress Remarks section (18-25) by reducing row heights
            for i in range(min(ot_extra, 8 - method_offset)):
                target_r = 18 + method_offset + i
                sheet.row_dimensions[target_r].height = 5
        
        base_shift = method_offset + ot_extra
        for i in range(2 + ot_extra):
            r = ot_header_row + 1 + i
            for col_let in ['B', 'F', 'I', 'K', 'N', 'S']:
                safe_write(f"{col_let}{r}", '')
        
        total_ot_hours = 0.0
        seen_ot_workers = set()
        for i, ot in enumerate(final_ot_list):
            r = ot_header_row + 1 + i
            original_names = ot.get('names', [])
            
            raw_h = ot.get('ot_hours', '')
            clean_h = _re.sub(r'\(.*?\)', '', raw_h).strip()
            try: h_val = float(clean_h)
            except: h_val = clean_h
            
            ot_amount = ot.get('ot_amount', '')
            
            # 개별 이름 단위로 중복 체크
            new_names = [n for n in original_names if n not in seen_ot_workers]
            
            if not new_names:
                # 모든 작업자가 이미 이전에 OT를 기록했으면 이름, 시간, 금액 모두 빈칸
                worker_name_display = ''
                h_val = ''
                ot_amount = ''
            else:
                # 아직 OT 기록이 안 된 작업자들만 이름 표시 및 OT 누적
                worker_name_display = ", ".join(new_names)
                for n in new_names:
                    seen_ot_workers.add(n)
                if isinstance(h_val, float):
                    total_ot_hours += h_val
            
            safe_write(f"B{r}", worker_name_display)
            safe_write(f"F{r}", ot.get('company', ''))
            safe_write(f"I{r}", ot.get('method', ''))
            safe_write(f"K{r}", h_val)
            safe_write(f"N{r}", ot_amount, is_currency=True)
            
        # 6. Section 6 (Materials)
        materials_data = data.get('materials', {}); active_rt = []
        for k, v in materials_data.items():
            if isinstance(v, dict) and (k.startswith('RT ') or k.startswith('RT_ROW_') or v.get('is_rt')):
                name = v.get('name', '').strip()
                if not name: continue
                item = v.copy(); item['name'] = name; active_rt.append(item)
        
        rt_count = len(active_rt)
        base_rt_limit = 3
        rt_extra = max(0, rt_count - base_rt_limit)
        
        if rt_extra > 0:
            sheet.insert_rows(46, rt_extra)
            for r_new in range(46, 46 + rt_extra):
                sheet.row_dimensions[r_new].height = 30
                for col in range(1, 20):
                    try:
                        source_cell = sheet.cell(row=45, column=col)
                        target_cell = sheet.cell(row=r_new, column=col)
                        if source_cell.has_style:
                            target_cell.font = copy(source_cell.font)
                            target_cell.border = copy(source_cell.border)
                            target_cell.fill = copy(source_cell.fill)
                            target_cell.alignment = copy(source_cell.alignment)
                            target_cell.number_format = source_cell.number_format
                    except: pass
        
        total_offset = base_shift + rt_extra
        def safe_merge(sheet, s_row, s_col, e_row, e_col, value):
            for m_range in list(sheet.merged_cells.ranges):
                if not (e_row < m_range.min_row or s_row > m_range.max_row or 
                        e_col < m_range.min_col or s_col > m_range.max_col):
                    try: sheet.unmerge_cells(m_range.coord)
                    except: pass
            try:
                sheet.merge_cells(start_row=s_row, start_column=s_col, end_row=e_row, end_column=e_col)
                cell = sheet.cell(row=s_row, column=s_col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
            except: pass

        rt_display_count = max(1, rt_count)
        rt_start_row = 43 + base_shift
        rt_end_row = (rt_start_row - 1) + rt_display_count
        
        safe_merge(sheet, rt_start_row, 2, rt_end_row, 3, "RT") 
        safe_merge(sheet, 46 + total_offset, 2, 47 + total_offset, 3, "MT")
        safe_merge(sheet, 48 + total_offset, 2, 50 + total_offset, 3, "PT")

        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 12

        for r in range(43 + base_shift, 52 + total_offset):
            try:
                for m_range in list(sheet.merged_cells.ranges):
                    if m_range.min_row == r and m_range.max_row == r and (m_range.min_col >= 4 and m_range.max_col <= 7):
                        sheet.unmerge_cells(m_range.coord)
                sheet.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
                sheet.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
                for s_c, e_c in [(9,10), (11,12), (13,14), (15,16), (17,19)]:
                    try:
                        for m_range in list(sheet.merged_cells.ranges):
                            if m_range.min_row == r and m_range.max_row == r and m_range.min_col == s_c:
                                sheet.unmerge_cells(m_range.coord)
                        sheet.merge_cells(start_row=r, start_column=s_c, end_row=r, end_column=e_c)
                    except: pass
            except: pass

        mat_start = 43 + base_shift
        for r in range(mat_start, mat_start + 4 + rt_extra):
            for col in range(4, 20):
                cell = sheet.cell(row=r, column=col); 
                if not isinstance(cell, MergedCell): cell.value = None

        def strip_brand(name):
            for prefix in ['Carestream ', 'AGFA ', 'Fuji ', 'Kodak ']:
                if name.startswith(prefix): return name[len(prefix):]
            return name

        for idx, m in enumerate(active_rt):
            r = mat_start + idx
            if r > 46 + rt_extra: break
            display_name = strip_brand(m.get('name', ''))
            safe_write(f'D{r}', display_name)
            cell_name = sheet.cell(row=r, column=4)
            cell_name.font = Font(name='맑은 고딕', size=9)
            cell_name.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            sheet.row_dimensions[r].height = 30
            safe_write(f'F{r}', m.get('spec', ''))
            sheet.cell(row=r, column=6).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            safe_write(f'H{r}', m.get('unit', '매'))
            safe_write(f'K{r}', '-'); safe_write(f'M{r}', f"{int(m.get('used', 0))} 매")
            safe_write(f'O{r}', f"{int(m.get('in', 0))} 매")

        for r_clear in range(18 + method_offset, 26 + method_offset):
            for c_clear in range(2, 20):
                cell = sheet.cell(row=r_clear, column=c_clear)
                if not isinstance(cell, MergedCell): cell.value = None

        tiny_font = Font(size=1); no_wrap = Alignment(wrap_text=False, horizontal='center', vertical='center')
        note_range_start = 18 + method_offset; note_range_end = 25 + method_offset
        for r in range(note_range_start, note_range_end + 1):
            sheet.row_dimensions[r].height = 6.0; sheet.row_dimensions[r].custom_height = True
            for c in range(1, 20):
                cell = sheet.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.value = None; cell.alignment = no_wrap; cell.font = tiny_font
        
        merges_to_kill = [m for m in list(sheet.merged_cells.ranges) if m.min_row >= note_range_start and m.max_row <= note_range_end]
        for m in merges_to_kill:
            try: sheet.unmerge_cells(str(m))
            except: pass
            
        for r in range(13, 18 + method_offset):
            sheet.row_dimensions[r].height = 15.0; sheet.row_dimensions[r].custom_height = True
        for r in range(43 + base_shift, 43 + base_shift + rt_count):
            sheet.row_dimensions[r].height = 25.0; sheet.row_dimensions[r].custom_height = True
        title_row = 41 + base_shift; header_row = 42 + base_shift
        sheet.row_dimensions[title_row].height = 20; sheet.row_dimensions[header_row].height = 20
        sheet.row_dimensions[title_row].custom_height = True; sheet.row_dimensions[header_row].custom_height = True

        if rt_count < 4:
            for r in range(43 + base_shift + rt_count, 46 + base_shift):
                sheet.row_dimensions[r].height = 0; sheet.row_dimensions[r].hidden = True
                for col in range(1, 20):
                    cell = sheet.cell(row=r, column=col)
                    if not isinstance(cell, MergedCell):
                        cell.value = None; cell.border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

        for r in range(43 + base_shift, 43 + base_shift + rt_count):
            if r >= mat_start + len(active_rt):
                safe_write(f'K{r}', '-'); safe_write(f'M{r}', '매'); safe_write(f'O{r}', '0 매')
            
        synonyms = {
            'MT WHITE': ['MT WHITE', '백색', '백색자분', 'WHITE'],
            'MT 7C-BLACK': ['MT 7C-BLACK', '7C', '자분', 'BLACK', '7C-BLACK'],
            'PT Penetrant': ['PT Penetrant', '침투', '침투액', 'Penetrant'],
            'PT Cleaner': ['PT Cleaner', '세척', '세척액', 'Cleaner'],
            'PT Developer': ['PT Developer', '현상', '현상액', '현상제', 'Developer']
        }
        mat_map = {
            'MT WHITE': 46 + total_offset, 'MT 7C-BLACK': 47 + total_offset, 
            'PT Penetrant': 48 + total_offset, 'PT Cleaner': 49 + total_offset, 'PT Developer': 50 + total_offset
        }
        display_names = {
            'MT WHITE': '백색페인트', 'MT 7C-BLACK': '흑색자분', 
            'PT Penetrant': '침투액', 'PT Cleaner': '세척액', 'PT Developer': '현상액'
        }
        
        for m_key, r in mat_map.items():
            display_name = display_names.get(m_key, m_key)
            safe_write(f'D{r}', display_name)
            try:
                for m_range in list(sheet.merged_cells.ranges):
                    if m_range.min_row == r and m_range.max_row == r and (m_range.min_col >= 4 and m_range.max_col <= 7):
                        sheet.unmerge_cells(m_range.coord)
                sheet.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
            except: pass
            safe_write(f'H{r}', "CAN")
            m_d = {}
            for syn in synonyms.get(m_key, []):
                m_d = data.get('materials', {}).get(syn)
                if m_d: break
            if m_d:
                safe_write(f'K{r}', '-'); safe_write(f'M{r}', int(m_d.get('used', 0)))
                safe_write(f'O{r}', int(m_d.get('in', 0)))

        thin = Side(style='thin')
        for r in range(41 + base_shift, 51 + total_offset):
            for col in range(2, 20):
                cell = sheet.cell(row=r, column=col)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = center_align
        
        # [CLEANUP] Explicitly remove borders from the row after the table
        for col in range(1, 21):
            sheet.cell(row=51 + total_offset, column=col).border = Border()

        for m_key, r in mat_map.items():
            display_name = display_names.get(m_key, m_key)
            cell = sheet.cell(row=r, column=4); cell.value = display_name
            cell.font = Font(name='맑은 고딕', size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            safe_write(f'H{r}', "CAN")
            m_d = {}
            for syn in synonyms.get(m_key, []):
                m_d = data.get('materials', {}).get(syn)
                if m_d: break
            if m_d:
                safe_write(f'K{r}', '-'); safe_write(f'M{r}', int(m_d.get('used', 0)))
                safe_write(f'O{r}', int(m_d.get('in', 0)))

        sheet.row_breaks.brk = []; sheet.print_area = f'A1:S{51 + total_offset}'
        from openpyxl.worksheet.pagebreak import Break
        sheet.row_breaks.append(Break(id=51 + total_offset))
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToHeight = None; sheet.page_setup.fitToWidth = None
        sheet.sheet_properties.pageSetUpPr.fitToPage = False

        veh_list = data.get('vehicles', []); v = veh_list[0] if veh_list else {}
        veh_row = 27 + method_offset; safe_write(f"B{veh_row}", "") 
        chk_rows = {'out': 29 + total_offset, 'in': 30 + total_offset}
        for rs in range(25+total_offset, 35+total_offset):
            val = sheet.cell(row=rs, column=2).value
            if val == "출차시": chk_rows['out'] = rs
            if val == "입차시": chk_rows['in'] = rs
        
        chk_map = {'exterior':'E','cleanliness':'H','cleaning':'K','locking':'N'}
        WHITE_SQ, BLACK_SQ = "\u25a1", "\u25a0"
        import re
        for rk, row_idx in chk_rows.items():
            for ck, col_let in chk_map.items():
                val = v.get(f"{rk}_{ck}")
                if val:
                    coord = f"{col_let}{row_idx}"; cell = sheet[coord]
                    if isinstance(cell, MergedCell):
                        for mr in sheet.merged_cells.ranges:
                            if coord in mr: cell = sheet.cell(row=mr.min_row, column=mr.min_col); break
                    if cell.value and isinstance(cell.value, str):
                        pattern = f"({WHITE_SQ})(\\s*){re.escape(val)}"
                        if re.search(pattern, cell.value):
                            cell.value = re.sub(pattern, f"{BLACK_SQ}\\2{val}", cell.value)

        sheet.page_setup.horizontalCentered = True; sheet.page_setup.verticalCentered = False 
        if v.get('remarks'):
            rem_cell = f"B{31 + method_offset}"
            safe_write(rem_cell, f"비고: {v.get('remarks')}")
            if sheet[rem_cell].value:
                sheet[rem_cell].alignment = Alignment(horizontal='left', vertical='center', indent=1)
                sheet[rem_cell].font = Font(name='맑은 고딕', size=9)

        sheet.page_margins.top = 0.4; sheet.page_margins.bottom = 0.4
        sheet.page_margins.left = 0.8; sheet.page_margins.right = 0.2
        for r in range(1, 12): sheet.row_dimensions[r].height = 15
        for t_row in [12, 26, 31, 37, 42]:
            try: sheet.row_dimensions[t_row + method_offset].height = 20
            except: pass
        for v_row in [27, 28, 29, 30]: 
            try: sheet.row_dimensions[v_row + method_offset].height = 32
            except: pass
        for r in range(41 + base_shift, 52 + total_offset): sheet.row_dimensions[r].height = 18
        compact_height = 10
        for r in [5, 10, 17, 26 + method_offset, 31 + method_offset, 35 + method_offset, 40 + base_shift]:
            try: sheet.row_dimensions[r].height = compact_height
            except: pass
            
        # 8. Final Print Setup - FIXED SCALE 95%
        sheet.sheet_properties.pageSetUpPr.fitToPage = False
        sheet.page_setup.scale = 95
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.horizontalCentered = True
        sheet.page_setup.verticalCentered = False
        sheet.page_margins.top = 0.6; sheet.page_margins.bottom = 0.1
        sheet.print_area = "A1:S51"
        
        rt_start_final = 43 + base_shift; rt_end_final = (rt_start_final - 1) + max(1, rt_count)
        thin_side = Side(style='thin'); thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        title_row_idx = 41 + base_shift; header_row_idx = 42 + base_shift
        for r_fix in [title_row_idx, header_row_idx]:
            for c_idx in range(1, 20): free_cell(sheet, r_fix, c_idx)
        write_cell(sheet, title_row_idx, 1, "6.자재수불현황", Font(name='맑은 고딕', size=9, bold=True), Alignment(horizontal='left', vertical='center'))
        header_merges = [(2,3, "구분"), (4,5, "품명"), (6,7, "규격"), (8,8, "단위"), (9,10, "Lot. No."), (11,12, "입고"), (13,14, "사용"), (15,16, "잔고"), (17,19, "비고")]
        header_font = Font(name='맑은 고딕', size=9, bold=True)
        for sc, ec, label in header_merges:
            if sc != ec: sheet.merge_cells(start_row=header_row_idx, start_column=sc, end_row=header_row_idx, end_column=ec)
            write_cell(sheet, header_row_idx, sc, label, header_font, center_align)
            for c_idx in range(sc, ec + 1): sheet.cell(row=header_row_idx, column=c_idx).border = thin_border
        
        merge_plan = [
            (rt_start_final, 2, rt_end_final, 3, "RT"), 
            (46 + total_offset, 2, 47 + total_offset, 3, "MT"), 
            (48 + total_offset, 2, 50 + total_offset, 3, "PT")
        ]
        for s_row, s_col, e_row, e_col, label in merge_plan:
            for r in range(s_row, e_row + 1):
                for c in range(s_col, e_col + 1): free_cell(sheet, r, c)
            try: sheet.merge_cells(start_row=s_row, start_column=s_col, end_row=e_row, end_column=e_col)
            except: pass
            write_cell(sheet, s_row, s_col, label, black_font, center_align)
            for r in range(s_row, e_row + 1):
                for c in range(s_col, e_col + 1):
                    key = (r, c)
                    if key not in sheet._cells or isinstance(sheet._cells[key], MergedCell): sheet._cells[key] = Cell(sheet, row=r, column=c)
                    sheet._cells[key].border = thin_border
        
        for m_key, r in mat_map.items():
            write_cell(sheet, r, 4, display_names.get(m_key, m_key), black_font, center_align)

        def apply_section_style(sheet, s_row, s_col, e_row, e_col, b_style='thin', inner_style='hair'):
            main_s = Side(style=b_style); inner_side = Side(style=inner_style) if inner_style else Side(style=None)
            for r in range(s_row, e_row + 1):
                for c in range(s_col, e_col + 1):
                    cell = sheet.cell(row=r, column=c)
                    t = inner_side; b = inner_side; l = inner_side; ri = inner_side
                    if r == s_row: t = main_s
                    if r == e_row: b = main_s
                    if c == s_col: l = main_s
                    if c == e_col: ri = main_s
                    cell.border = Border(top=t, bottom=b, left=l, right=ri)

        s2_start = 18 + method_offset; s2_end = 25 + method_offset
        for m_range in list(sheet.merged_cells.ranges):
            if s2_start <= m_range.min_row and m_range.max_row <= s2_end:
                try: sheet.unmerge_cells(m_range.coord)
                except: pass
        for r_idx in range(s2_start, s2_end + 1):
            try: sheet.merge_cells(start_row=r_idx, start_column=2, end_row=r_idx, end_column=19)
            except: pass
        apply_section_style(sheet, s2_start, 2, s2_end, 19, 'thin', inner_style=None)
        apply_section_style(sheet, 12, 2, 16 + method_offset, 19, 'thin')
        apply_section_style(sheet, 6, 2, 9, 19, 'thin')
        v_title_row = 26 + method_offset; v_data_start = 28 + method_offset; v_data_end = 30 + method_offset
        apply_section_style(sheet, v_data_start, 2, v_data_end, 19, 'thin')
        for c_idx in range(2, 20): sheet.cell(row=v_title_row, column=c_idx).border = Border(top=thin_side) 
        
        # 31행 좌/우/하단 선 제거 (상단선만 유지하거나 테두리 전체 제거)
        v_remark_row = 31 + method_offset
        for c_idx in range(2, 20):
            c = sheet.cell(row=v_remark_row, column=c_idx)
            # 기존 위쪽 선은 유지하고 좌,우,아래는 없앰
            top_border = c.border.top if c.border else Side(style=None)
            c.border = Border(top=top_border, bottom=Side(style=None), left=Side(style=None), right=Side(style=None))
        apply_section_style(sheet, ot_header_row, 2, ot_header_row + 2 + ot_extra, 19, 'thin')
        apply_section_style(sheet, 43 + base_shift, 2, 50 + total_offset, 19, 'thin')
        for c in range(1, 20):
            # 자재수불현황 제목행(41+base_shift)은 윗선을 포함한 모든 선을 완전히 제거
            sheet.cell(row=41 + base_shift, column=c).border = Border()
        if rtk_header_row: apply_section_style(sheet, rtk_header_row + 1, 2, rtk_header_row + 2, 19, 'thin')
        for r_idx in [3, 4]:
            c = sheet.cell(row=r_idx, column=19); c.border = Border(left=c.border.left, top=c.border.top, bottom=c.border.bottom, right=thin_side)
        c5 = sheet.cell(row=5, column=19); c5.border = Border(left=c5.border.left, top=c5.border.top, bottom=c5.border.bottom, right=Side(style=None))
        
        # [RESTORED] Column Widths Setup
        slim_widths = {
            'A': 1, 'B': 5, 'C': 1, 'D': 6, 'E': 6, 'F': 6, 'G': 1, 'H': 5, 
            'I': 4, 'J': 4, 'K': 5, 'L': 4, 'M': 5, 'N': 5, 'O': 4, 
            'P': 4, 'Q': 4, 'R': 4, 'S': 5
        }
        for col, width in slim_widths.items():
            sheet.column_dimensions[col].width = width
            
        wb.save(output_path); return output_path
