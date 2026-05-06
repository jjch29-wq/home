from openpyxl import load_workbook
import os

path = os.path.abspath("../resources/Template_DailyWorkReport.xlsx")
wb = load_workbook(path)
sheet = wb.active

for r in range(1, 100):
    row_text = "".join([str(sheet.cell(row=r, column=c).value or '') for c in range(1, 5)])
    if "5." in row_text:
        print(f"Row {r}: {row_text}")
        # Peek at data rows below it
        for i in range(1, 6):
            data_row = [sheet.cell(row=r+i, column=c).value for c in range(1, 10)]
            print(f"  Row {r+i}: {data_row}")

wb.close()
