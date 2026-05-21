import openpyxl

file_path = r'c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\Na-aba\home\data\SISGI-KOGAS-RT-001.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

print("Sheets in template:", wb.sheetnames)
for sheet_name in wb.sheetnames[:2]:
    ws = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    for r in range(1, 15):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 35)]
        # Filter out empty rows or print them nicely
        if any(row_vals):
            print(f"Row {r:2d}: {[str(v)[:15] if v is not None else '' for v in row_vals]}")

wb.close()
