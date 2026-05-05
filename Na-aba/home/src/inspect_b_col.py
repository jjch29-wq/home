from openpyxl import load_workbook
import os

path = os.path.abspath("../resources/Template_DailyWorkReport.xlsx")
wb = load_workbook(path)
sheet = wb.active

for r in [29, 30]:
    val = sheet.cell(row=r, column=2).value
    if val:
        print(f"B{r} hex: {' '.join(hex(ord(c)) for c in val)}")
        print(f"B{r} text: {val}")

wb.close()
