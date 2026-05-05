from openpyxl import load_workbook
import os

path = os.path.abspath("../resources/Template_DailyWorkReport.xlsx")
wb = load_workbook(path)
sheet = wb.active

for r in range(25, 30):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 6)]
    print(f"Row {r}: {vals}")

wb.close()
