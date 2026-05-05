from openpyxl import load_workbook
import os

path = os.path.abspath("../resources/Template_DailyWorkReport.xlsx")
wb = load_workbook(path)
sheet = wb.active

print("Peeking B26-B31:")
for r in range(26, 32):
    val = sheet.cell(row=r, column=2).value # Column B
    print(f"B{r}: {repr(val)}")

print("\nPeeking E26-E31:")
for r in range(26, 32):
    val = sheet.cell(row=r, column=5).value # Column E
    print(f"E{r}: {repr(val)}")

wb.close()
