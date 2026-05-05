from openpyxl import load_workbook
import os

path = os.path.abspath("../resources/Template_DailyWorkReport.xlsx")
wb = load_workbook(path)
sheet = wb.active

for r in [26, 27, 28, 29, 30]:
    print(f"\nRow {r} Analysis:")
    # Check merges
    for rng in sheet.merged_cells.ranges:
        if r >= rng.min_row and r <= rng.max_row:
            print(f"  Merge found: {rng}")
    
    # Check height
    height = sheet.row_dimensions[r].height
    print(f"  Height: {height}")
    
    # Check font
    cell = sheet.cell(row=r, column=1)
    cell_e = sheet.cell(row=r, column=5)
    if cell.value:
        print(f"  A{r} Value: {repr(cell.value)}")
    if cell_e.value:
        print(f"  E{r} Value: {repr(cell_e.value)}")

wb.close()
