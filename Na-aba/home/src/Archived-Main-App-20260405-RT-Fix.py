import sys
import os
import glob
import math
import traceback
import re
import warnings
import json
import tempfile
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import threading
import xlsxwriter
import datetime
from PIL import Image as PILImage, ImageChops, ImageOps
import io
import time

# DPI Awareness for Windows
try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except:
    pass

# --- Versioning ---
APP_VERSION = "v260404.03"
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
# Ignore library warnings
warnings.simplefilter("ignore")

# --- Constants & Paths ---
NAN_PATTERN = re.compile(r'^nan(\.0+)?$|^none$|^null$|^0\.0+|-0\.0+$', re.IGNORECASE)
DOT_ZERO_PATTERN = re.compile(r'\.0$')

# [PT] SCH -> 두께(mm) 변환 테이블
SCH_TO_THK = {
    "1/2": {"5S": 1.65, "10S": 2.11, "40": 2.77, "80": 3.73, "160": 4.75, "XXS": 7.47},
    "3/4": {"5S": 1.65, "10S": 2.11, "40": 2.87, "80": 3.91, "160": 5.56, "XXS": 7.82},
    "1": {"5S": 1.65, "10S": 2.77, "40": 3.38, "80": 4.55, "160": 6.35, "XXS": 9.09},
    "1-1/4": {"5S": 1.65, "10S": 2.77, "40": 3.56, "80": 4.85, "160": 6.35, "XXS": 9.70},
    "1-1/2": {"5S": 1.65, "10S": 2.77, "40": 3.68, "80": 5.08, "160": 7.14, "XXS": 10.16},
    "2": {"5S": 1.65, "10S": 2.77, "40": 3.91, "80": 5.54, "160": 8.74, "XXS": 11.07},
    "2-1/2": {"5S": 2.11, "10S": 3.05, "40": 5.16, "80": 7.01, "160": 9.53, "XXS": 14.02},
    "3": {"5S": 2.11, "10S": 3.05, "40": 5.49, "80": 7.62, "160": 11.13, "XXS": 15.24},
    "4": {"5S": 2.11, "10S": 3.05, "40": 6.02, "80": 8.56, "160": 13.49, "XXS": 17.12},
    "5": {"5S": 2.77, "10S": 3.40, "40": 6.55, "80": 9.53, "120": 12.70, "160": 15.88},
    "6": {"5S": 2.77, "10S": 3.40, "40": 7.11, "80": 10.97, "120": 14.27, "160": 18.26, "XXS": 21.95},
    "8": {"5S": 2.77, "10S": 3.76, "20": 6.35, "30": 7.04, "40": 8.18, "60": 10.31, "80": 12.70, "100": 15.09, "120": 18.26, "140": 20.62, "160": 23.01, "XXS": 22.23},
    "10": {"5S": 3.40, "10S": 4.19, "20": 6.35, "30": 7.80, "40": 9.27, "60": 12.70, "80": 15.09, "100": 18.26, "120": 21.44, "140": 25.40, "160": 28.58},
    "12": {"5S": 3.96, "10S": 4.57, "20": 6.35, "30": 8.38, "40": 10.31, "60": 14.27, "80": 17.48, "100": 21.44, "120": 25.40, "140": 28.58, "160": 33.32},
    "14": {"5S": 3.96, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 9.53, "40": 11.13, "60": 15.09, "80": 19.05, "100": 23.83, "120": 27.79, "140": 31.75, "160": 35.71},
    "16": {"5S": 4.19, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 9.53, "40": 12.70, "60": 16.66, "80": 21.44, "100": 26.19, "120": 30.96, "140": 36.53, "160": 40.49},
    "18": {"5S": 4.19, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 11.13, "40": 14.27, "60": 19.05, "80": 23.83, "100": 29.36, "120": 34.93, "140": 39.67, "160": 45.24},
    "20": {"5S": 4.78, "10S": 5.54, "10": 6.35, "20": 9.53, "30": 12.70, "40": 15.09, "60": 20.62, "80": 26.19, "100": 32.54, "120": 38.10, "140": 44.45, "160": 50.01},
    "24": {"5S": 5.54, "10S": 6.35, "10": 6.35, "20": 9.53, "30": 14.27, "40": 17.48, "60": 24.61, "80": 30.96, "100": 38.89, "120": 46.02, "140": 52.37, "160": 59.54},
}

def convert_sch_to_thk(size_val, thk_val):
    """SCH 값을 두께(mm)로 변환"""
    if pd.isna(thk_val) or str(thk_val).strip() == "": return ""
    thk_str = str(thk_val).strip().upper()
    try:
        val = float(thk_str.replace("MM", "").replace("T", "").strip())
        if 0 < val < 100: return f"{val:.2f}"
    except: pass
    sch_match = re.search(r'(?:SCH[.\s]?|S/)?(\d+S?|XXS|XS)', thk_str, re.IGNORECASE)
    if not sch_match: return thk_str
    sch = sch_match.group(1).upper()
    if sch.endswith('S') and sch not in ['5S', '10S', 'XXS', 'XS']: sch = sch[:-1]
    if pd.isna(size_val) or str(size_val).strip() == "": return thk_str
    size_str = str(size_val).strip().replace('"', '').replace("'", "")
    size_str = re.sub(r'\s+', '-', size_str)
    if size_str in SCH_TO_THK and sch in SCH_TO_THK[size_str]:
        return f"{SCH_TO_THK[size_str][sch]:.2f}"
    try:
        size_int = str(int(float(size_str)))
        if size_int in SCH_TO_THK and sch in SCH_TO_THK[size_int]:
            return f"{SCH_TO_THK[size_int][sch]:.2f}"
    except: pass
    return thk_str

# 공통 스타일 (테두리 등)
thin_side = Side(style='thin')
medium_side = Side(style='medium')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

if getattr(sys, 'frozen', False):
    # PyInstaller creates a temp folder and stores path in _MEIPASS
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
    SCRIPT_HOME = base_path
    BASE_DIR = base_path
    CONFIG_DIR = base_path
    RESOURCE_DIR = os.path.join(base_path, "resources")
else:
    SCRIPT_HOME = os.path.dirname(os.path.abspath(__file__))
    BASE_DIR = os.path.dirname(SCRIPT_HOME)
    CONFIG_DIR = os.path.join(BASE_DIR, "config")
    RESOURCE_DIR = os.path.join(BASE_DIR, "resources")

SETTINGS_FILE = os.path.join(CONFIG_DIR, "logo_settings_unified.json")

class PMIReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"SITCO 통합 성적서 자동 생성기 ({APP_VERSION})")
        self.root.geometry("950x850") # Slightly wider/taller for multi-tab
        self.root.configure(background="#f9fafb")
        
        # 1. Initialize Configuration first (needed for state variables)
        self.config = {
            # UI State (Persistence)
            'PMI_SASH_RATIO': 0.5, 'RT_SASH_RATIO': 0.5, 'PT_SASH_RATIO': 0.5, 'PAUT_SASH_RATIO': 0.5,
            'PMI_COL_WIDTHS': {}, 'RT_COL_WIDTHS': {}, 'PT_COL_WIDTHS': {}, 'PAUT_COL_WIDTHS': {},
            # 갑지 (Cover)
            'SEOUL_COVER_ANCHOR': "E5", 'SEOUL_COVER_W': 200.0, 'SEOUL_COVER_H': 18.0, 'SEOUL_COVER_X': 30.0, 'SEOUL_COVER_Y': 15.0,
            'SITCO_COVER_ANCHOR': "A6", 'SITCO_COVER_W': 80.0, 'SITCO_COVER_H': 40.0, 'SITCO_COVER_X': 10.0, 'SITCO_COVER_Y': 10.0,
            'FOOTER_COVER_ANCHOR': "P50", 'FOOTER_COVER_W': 80.0, 'FOOTER_COVER_H': 20.0, 'FOOTER_COVER_X': 0.0, 'FOOTER_COVER_Y': 10.0,
            'FOOTER_PT_COVER_ANCHOR': "A50", 'FOOTER_PT_COVER_W': 100.0, 'FOOTER_PT_COVER_H': 25.0, 'FOOTER_PT_COVER_X': 5.0, 'FOOTER_PT_COVER_Y': 10.0,
            # 을지 (Data)
            'SEOUL_DATA_ANCHOR': "F5", 'SEOUL_DATA_W': 200.0, 'SEOUL_DATA_H': 18.0, 'SEOUL_DATA_X': 30.0, 'SEOUL_DATA_Y': 15.0,
            'SITCO_DATA_ANCHOR': "A6", 'SITCO_DATA_W': 100.0, 'SITCO_DATA_H': 50.0, 'SITCO_DATA_X': 20.0, 'SITCO_DATA_Y': 5.0,
            'FOOTER_DATA_ANCHOR': "M46", 'FOOTER_DATA_W': 100.0, 'FOOTER_DATA_H': 15.0, 'FOOTER_DATA_X': 25.0, 'FOOTER_DATA_Y': 10.0,
            'FOOTER_PT_DATA_ANCHOR': "A46", 'FOOTER_PT_DATA_W': 100.0, 'FOOTER_PT_DATA_H': 25.0, 'FOOTER_PT_DATA_X': 3.0, 'FOOTER_PT_DATA_Y': 7.0,
            'MARGIN_DATA_TOP': 0.2, 'MARGIN_DATA_BOTTOM': 0.0, 'MARGIN_DATA_LEFT': 0.5, 'MARGIN_DATA_RIGHT': 0.3, 'PRINT_SCALE_DATA': 95,
            'START_ROW': 19, 'DATA_END_ROW': 45, 'PRINT_END_ROW': 47,
            # [NEW] 선택적 행/열 조절용 설정
            'CUSTOM_ROWS_COVER': '23-38', 'CUSTOM_ROW_HEIGHT_COVER': 21.0,
            'CUSTOM_COLS_COVER': '', 'CUSTOM_COL_WIDTH_COVER': 10.0,
            'CUSTOM_ROWS_DATA': '', 'CUSTOM_ROW_HEIGHT_DATA': 20.55,
            'CUSTOM_COLS_DATA': '', 'CUSTOM_COL_WIDTH_DATA': 10.0,
            
            # PAUT Specific (B31.1)
            'PAUT_START_ROW': 11, 'PAUT_DATA_END_ROW': 40, 'PAUT_PRINT_END_ROW': 45,
            
            # RT Specific (Radiographic Testing)
            'RT_START_ROW': 11, 'RT_DATA_END_ROW': 45, 'RT_PRINT_END_ROW': 47,
            
            # Photo Log Sash
            'PHOTO_SASH_RATIO': 0.45
        }

        # 2. State Variables
        self.setting_vars = {} # [CRITICAL FIX] Container for all dynamic UI variables
        self.logo_folder_path = tk.StringVar(value="")
        self.target_file_path = tk.StringVar(value=self.config.get('PMI_TARGET_PATH', ""))
        self.template_file_path = tk.StringVar(value=self.config.get('PMI_TEMPLATE_PATH', ""))
        self.sequence_filter = tk.StringVar()
        self.rt_extract_keyword = tk.StringVar()  # [NEW] 추출 키워드 필터
        self.extraction_mode = tk.StringVar(value="전체")
        self.auto_verify = tk.BooleanVar(value=True)
        self.pmi_pane_ratio = self.config.get('PMI_SASH_RATIO', 0.5)
        self.current_mode = "PMI" # [NEW] Track active report mode
        
        self.show_selected_only = tk.BooleanVar(value=False)
        self.extracted_data = [] 
        self.pmi_sort_col = "" 
        self.pmi_sort_rev = False 
        self.rt_sort_col = ""
        self.rt_sort_rev = False
        self.pt_sort_col = ""
        self.pt_sort_rev = False
        self.paut_sort_col = ""
        self.paut_sort_rev = False
        
        # [NEW] PMI Preview Search & Filter
        self.pmi_search_loc = tk.StringVar()
        self.pmi_show_deficiency_only = tk.BooleanVar(value=False)
        
        # [NEW] 원소 함량 필터링용 상태 변수 (PMI 전용)
        self.element_filters = [] # list of dict: {'key': StringVar, 'min': StringVar, 'max': StringVar}
        # 기본 필터 추가 (Cr, Ni, Mo)
        for k in ["Cr", "Ni", "Mo"]:
            self.element_filters.append({
                'key': tk.StringVar(value=k), 
                'min': tk.StringVar(value=''), 
                'max': tk.StringVar(value='')
            })
            
        # --- PAUT State Variables ---
        self.paut_target_file_path = tk.StringVar(value=self.config.get('PAUT_TARGET_PATH', ""))
        self.paut_template_file_path = tk.StringVar(value=self.config.get('PAUT_TEMPLATE_PATH', ""))
        self.paut_manual_vars = {
            't': tk.StringVar(), 'h': tk.StringVar(), 'l': tk.StringVar(), 'd': tk.StringVar(),
            'nature': tk.StringVar(value="Slag"), 'loc': tk.StringVar(value="-"),
            'db': tk.StringVar(value="6"), 'peak': tk.StringVar(value="80"), 
            'z1': tk.StringVar(), 'z2': tk.StringVar(), 
            'L1': tk.StringVar(), 'L2': tk.StringVar(),
            'D_Upper': tk.StringVar(), 'D_Lower': tk.StringVar(),
            'target_fsh': tk.StringVar(value="40.0%")
        }
        self.paut_eval_mode = tk.StringVar(value="ECA") # [NEW] Standard / ECA
        self.paut_manual_vars['peak'].trace_add("write", lambda *args: self._update_paut_target_fsh())
        self.paut_manual_vars['db'].trace_add("write", lambda *args: self._update_paut_target_fsh())
        self.paut_extracted_data = []
        
        # --- RT State Variables ---
        self.rt_target_file_path = tk.StringVar(value=self.config.get('RT_TARGET_PATH', ""))
        self.rt_template_file_path = tk.StringVar(value=self.config.get('RT_TEMPLATE_PATH', ""))
        self.kogas_target_file_path = tk.StringVar()
        self.kogas_template_file_path = tk.StringVar()
        self.rt_kogas_mode = tk.BooleanVar(value=False) 
        self.rt_kogas_mode.trace_add("write", lambda *a: self._on_rt_kogas_mode_change())
        self.extracted_data = []
        self.rt_extracted_data = []
        self.kogas_extracted_data = []
        
        # --- Column Keys Initialization ---
        self.column_keys = ["selected", "No", "Date", "Dwg", "Joint", "Loc", "Ni", "Cr", "Mo", "Result"]
        self.rt_column_keys = ["selected", "No", "Date", "Dwg", "Joint", "Loc", "T", "Mat", "Deg", "Acc", "Rej", "D1", "D2", "D3", "D4", "D5", "D6", "D7", "D8", "D9", "D10", "D11", "D12", "D13", "D14", "D15", "Result", "Welder", "Remarks"]
        self.kogas_column_keys = self.rt_column_keys
        self.pt_column_keys = ["selected", "No", "Date", "Dwg", "Joint", "Loc", "T", "Mat", "Deg", "Result", "Welder", "Remarks"]
        self.paut_column_keys = ["selected", "No", "Date", "ISO", "Joint", "Loc", "T", "Mat", "Grade", "Nature", "Type", "a/l", "a/t", "Evaluation", "Remarks"]
        
        # --- PT State Variables ---
        self.pt_target_file_path = tk.StringVar(value=self.config.get('PT_TARGET_PATH', ""))
        self.pt_template_file_path = tk.StringVar(value=self.config.get('PT_TEMPLATE_PATH', ""))
        self.pt_extracted_data = []
        self.item_idx_map = []      # [NEW] Missing for PMI
        self.rt_item_idx_map = []   # [NEW] Missing for RT
        self.pt_item_idx_map = []
        self.paut_item_idx_map = [] 
        self.kogas_item_idx_map = []
        
        # [REFINED] Column Keys Mapping (Must match Treeview column count and order)
        self.column_keys = ["_status", "selected", "No", "Date", "Dwg", "Joint", "Loc", "Ni", "Cr", "Mo", "Grade"]
        self.rt_column_keys = ["selected", "No", "Date", "Dwg", "Joint", "Loc", "T", "Mat", "Acc", "Rej", "Deg", "D1", "D2", "D3", "D4", "D5", "D6", "D7", "D8", "D9", "D10", "D11", "D12", "D13", "D14", "D15", "Welder", "Remarks"]
        self.kogas_column_keys = self.rt_column_keys
        self.pt_column_keys = ["selected", "No", "Date", "Dwg", "Joint", "Material", "TestItem", "Result", "Welder", "Remarks"]
        self.paut_column_keys = ["selected", "No", "Line No.", "Joint No.", "Th'k(mm)", "Start", "End", "Length(mm)", "Upper", "Lower", "Height(mm)", "Type of Flaw", "a/l", "a/t", "Evaluation", "Remarks"]
        
        self.date_listbox = None
    
        # [NEW] Handle Application Closing for Final State Capture
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.rt_date_listbox = None
        self.paut_date_listbox = None
        self.kogas_date_listbox = None

        # [NEW] Add traces for Template-Linked Config Auto-Load
        self.rt_template_file_path.trace_add("write", lambda *args: self.load_template_specific_config(self.rt_template_file_path.get(), "RT"))
        self.kogas_template_file_path.trace_add("write", lambda *args: self.load_template_specific_config(self.kogas_template_file_path.get(), "KOGAS"))
        self.pt_template_file_path.trace_add("write", lambda *args: self.load_template_specific_config(self.pt_template_file_path.get(), "PT"))
        self.paut_template_file_path.trace_add("write", lambda *args: self.load_template_specific_config(self.paut_template_file_path.get(), "PAUT"))
        self.template_file_path.trace_add("write", lambda *args: self.load_template_specific_config(self.template_file_path.get(), "PMI"))

        # Keys are now verified to match Treeview column indices 1-N.

        # 3. UI Initialization
        # [NEW] File Info Strings for Preview Headers
        self.file_info_vars = {
            'PMI': tk.StringVar(value="📄 파일을 선택해주세요."),
            'RT': tk.StringVar(value="📄 파일을 선택해주세요."),
            'KOGAS': tk.StringVar(value="📄 파일을 선택해주세요."),
            'PT': tk.StringVar(value="📄 파일을 선택해주세요."),
            'PAUT': tk.StringVar(value="📄 파일을 선택해주세요."),
            'PHOTO': tk.StringVar(value="📸 사진 리스트를 구성해주세요.")
        }

        # --- Gapji (Cover) Metadata Variables ---
        self.gapji_project = tk.StringVar(value=self.config.get('GAPJI_PROJECT', ""))
        self.gapji_customer = tk.StringVar(value=self.config.get('GAPJI_CUSTOMER', ""))
        self.gapji_item = tk.StringVar(value=self.config.get('GAPJI_ITEM', ""))
        self.gapji_material = tk.StringVar(value=self.config.get('GAPJI_MATERIAL', ""))
        self.gapji_report_no = tk.StringVar(value=self.config.get('GAPJI_REPORT_NO', ""))
        self.gapji_exam_date = tk.StringVar(value=self.config.get('GAPJI_EXAM_DATE', datetime.datetime.now().strftime("%Y-%m-%d")))
        
        # [NEW] Trace Gapji metadata for real-time preview updates
        for var in [self.gapji_project, self.gapji_customer, self.gapji_item, 
                    self.gapji_material, self.gapji_report_no, self.gapji_exam_date]:
            var.trace_add("write", lambda *a: self._update_gapji_preview(getattr(self, 'current_mode', 'PMI')))

        # --- Photo Log State Variables ---
        self.photo_header_map = {
            "PAUT": "REPORT OF PHASED ARRAY UT EXAMINATION (위 상 배 열 초 음 파 탐 상 검 사 보 고 서)",
            "RT": "REPORT OF RADIOGRAPHIC EXAMINATION (방 사 선 투 과 검 사 보 고 서)",
            "PT": "REPORT OF LIQUID PENETRANT EXAMINATION (침 투 탐 상 검 사 보 고 서)",
            "MT": "REPORT OF MAGNETIC PARTICLE EXAMINATION (자 분 탐 상 검 사 보 고 서)",
            "PMI": "REPORT OF POSITIVE MATERIAL IDENTIFICATION (재 질 성 분 분 석 검 사 보 고 서)",
            "UT": "REPORT OF ULTRASONIC EXAMINATION (초 음 파 탐 상 검 사 보 고 서)",
            "NDT": "REPORT OF NON-DESTRUCTIVE EXAMINATION (비 파 괴 검 사 보 고 서)",
            "기타 (직접 입력)": ""
        }
        self.photo_orderer = tk.StringVar(value="서울에너지공사")
        self.photo_inspect_date = tk.StringVar(value=datetime.datetime.now().strftime("%Y-%m-%d"))
        self.photo_report_no = tk.StringVar(value="SIT/GI-SE-PAUT-TNTFJPWJ001")
        self.photo_inspect_type = tk.StringVar(value="PAUT")
        self.photo_report_title = tk.StringVar(value=self.photo_header_map["PAUT"])
        self.photo_cols_per_row = tk.StringVar(value="2")
        self.photo_keep_aspect = tk.BooleanVar(value=True)
        self.photo_output_name = tk.StringVar(value="NDT_Photo_Log_Final.xlsx")
        _def_logo = os.path.join(RESOURCE_DIR, "logo.png")
        self.photo_logo_path = tk.StringVar(value=_def_logo if os.path.exists(_def_logo) else "")
        self.photo_logo_width_var = tk.StringVar(value="80")
        self.photo_logo_x_var = tk.StringVar(value="2")
        self.photo_logo_y_var = tk.StringVar(value="0")
        self.photo_cell_width_var = tk.StringVar(value="53.0")
        self.photo_cell_height_var = tk.StringVar(value="178.0")
        self.photo_margin_top_var = tk.StringVar(value="0.5")
        self.photo_margin_bottom_var = tk.StringVar(value="0.5")
        self.photo_margin_left_var = tk.StringVar(value="0.4")
        self.photo_margin_right_var = tk.StringVar(value="0.4")
        self.photo_print_scale_var = tk.StringVar(value="100")
        self.photo_desc_height_var = tk.StringVar(value="20.0")
        self.photo_align_var = tk.StringVar(value="중앙 정렬")
        self.photo_fit_width_var = tk.BooleanVar(value=True)
        self.photo_auto_rotate_var = tk.BooleanVar(value=False)
        self.photo_width_pct_var = tk.StringVar(value="100.0")
        self.photo_width_pixel_adj_var = tk.StringVar(value="0")
        self.photo_shift_x_var = tk.StringVar(value="0")
        self.photo_shift_y_var = tk.StringVar(value="0")
        self.photo_selected_files = [] 
        
        self.load_settings()
        self.last_photo_save_dir = self.config.get('PHOTO_LOG_SETTINGS', {}).get('last_save_dir', "")

        self.create_widgets()
        
        # [NEW] Sync initial file info from loaded config
        self._sync_all_file_infos()
        
        # [NEW] Auto-save on exit
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        self.log("[INFO] 통합 버전을 시작했습니다.")
        
    def _safe_update_scrollregion(self):
        """Safely update canvas scrollregion, handling None bbox."""
        try:
            bbox = self.canvas.bbox("all")
            if bbox:
                self.canvas.configure(scrollregion=bbox)
        except Exception:
            pass

    def log(self, message):
        if hasattr(self, 'status_log'):
            self.status_log.config(state='normal')
            self.status_log.insert(tk.END, f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}\n")
            self.status_log.see(tk.END)
            self.status_log.config(state='disabled')
            self.root.update_idletasks()
        else:
            print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}")

    def fix_material_name(self, t):
        """재질명을 표준 포맷으로 변환"""
        t_str = str(t).upper()
        if t_str == 'NAN': return ""
        t_str = t_str.replace('A312-TP304', 'S/S').replace('A312-304L', 'S/S').replace('A312-305L', 'S/S').replace('A53-B', 'C/S').replace('A106-B', 'C/S')
        return t_str.replace('C2','C/S').replace('C4','C/S').replace('CS','C/S').replace('S99','S/S').replace('SS','S/S')

    def force_two_digit(self, val):
        """숫자 값을 2자리 문자열로 변환 (예: 1 -> 01)"""
        try:
            s = str(val).strip()
            f = float(s)
            if f.is_integer(): return f"{int(f):02d}"
            return s
        except: return str(val).strip()

    def load_settings(self):
        # 1. 내장 설정 파일 (실행파일 내부에 번들링된 기본값)
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            bundle_settings = os.path.join(sys._MEIPASS, "logo_settings_unified.json")
            if os.path.exists(bundle_settings):
                try:
                    with open(bundle_settings, 'r', encoding='utf-8') as f:
                        bundle_data = json.load(f)
                        self.config.update(bundle_data)
                    print("SUCCESS: 내장된 기본 설정을 로드했습니다.")
                except Exception as e:
                    print(f"WARNING: 내장 설정 불러오기 실패: {e}")

        # 2. 외부 설정 파일 (사용자 저장값 - 내장값보다 우선순위 높음)
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    saved_data = json.load(f)
                    
                    # [REFINED] Case-insensitive Sanitization for all modules
                    sanit_map = {
                        'PMI_NAME_DWG': (['ISO/DWG', 'ISO/DWG.'], 'Drawing No.'),
                        'PMI_NAME_JOINT': (['JOINT NO', 'JOINT NO.'], 'Joint No.'),
                        'PMI_NAME_LOC': (['TEST LOCATION', 'LOC'], 'Location'),
                        'PMI_NAME_RES': (['GRADE', 'RESULT'], 'Result'),
                        'RT_NAME_DWG': (['ISO/DWG', 'ISO/DWG.', 'DWG'], 'Drawing No.'),
                        'RT_NAME_JOINT': (['JOINT NO', 'FILM IDENT', 'JOINT'], 'Film Ident. No.'),
                        'RT_NAME_LOC': (['LOC', 'LOC.'], 'Film Location'),
                        'RT_NAME_WELDER': (['WELDER'], 'Welder No'),
                        'RT_NAME_DEG': (['물성', '물성(DEG)', 'DEG'], 'Deg')
                    }
                    for k, (olds, new) in sanit_map.items():
                        if k in saved_data:
                            val = str(saved_data[k]).upper().strip()
                            if any(val == o.upper() for o in olds):
                                saved_data[k] = new

                    # [NEW] Sanitize old "Element-" prefixes and Mn exclusion
                    for k in ['PMI_NAME_NI', 'PMI_NAME_CR', 'PMI_NAME_MO']:
                        if k in saved_data and isinstance(saved_data[k], str) and (saved_data[k].startswith("Element-") or "ELEMENT" in saved_data[k].upper()):
                            saved_data[k] = saved_data[k].replace("Element-", "").replace("ELEMENT-", "").replace("Element", "").strip()
                    
                    if 'column_keys' in saved_data and isinstance(saved_data['column_keys'], list):
                        while 'Mn' in saved_data['column_keys']: saved_data['column_keys'].remove('Mn')

                    # [NEW] Defect Names Sanitization (D1-D15 to circled labels)
                    defect_standards = {
                        1: "① Crack", 2: "② IP", 3: "③ LF", 4: "④ Slag", 5: "⑤ Por",
                        6: "⑥ U/C", 7: "⑦ RUC", 8: "⑧ BT", 9: "⑨ TI", 10: "⑩ CP",
                        11: "⑪ RC", 12: "⑫ Mis", 13: "⑬ EP", 14: "⑭ SD", 15: "⑮ Oth"
                    }
                    for d_i, std_name in defect_standards.items():
                        d_k = f'RT_NAME_D{d_i}'
                        if d_k in saved_data:
                            curr_val = str(saved_data[d_k]).upper().strip()
                            if curr_val == f"D{d_i}" or not curr_val or curr_val == "NAN":
                                saved_data[d_k] = std_name
                    
                    # [NEW] Restore Photo Log Variables
                    photo_vars = {
                        'orderer': self.photo_orderer, 'inspect_date': self.photo_inspect_date,
                        'inspect_type': self.photo_inspect_type, 'report_title': self.photo_report_title,
                        'report_no': self.photo_report_no,
                        'cols_per_row': self.photo_cols_per_row, 'keep_aspect': self.photo_keep_aspect,
                        'output_name': self.photo_output_name, 'logo_path': self.photo_logo_path,
                        'logo_width': self.photo_logo_width_var,
                        'logo_x': self.photo_logo_x_var, 'logo_y': self.photo_logo_y_var,
                        'cell_width': self.photo_cell_width_var, 'cell_height': self.photo_cell_height_var,
                        'm_top': self.photo_margin_top_var, 'm_bottom': self.photo_margin_bottom_var,
                        'm_left': self.photo_margin_left_var, 'm_right': self.photo_margin_right_var,
                        'print_scale': self.photo_print_scale_var, 'desc_height': self.photo_desc_height_var,
                        'photo_align': self.photo_align_var, 'fit_width': self.photo_fit_width_var,
                        'auto_rotate': self.photo_auto_rotate_var, 'width_pct': self.photo_width_pct_var,
                        'pixel_adj': self.photo_width_pixel_adj_var, 'shift_x': self.photo_shift_x_var,
                        'shift_y': self.photo_shift_y_var
                    }
                    if 'PHOTO_LOG_SETTINGS' in saved_data:
                        plist = saved_data['PHOTO_LOG_SETTINGS']
                        for pk, pvar in photo_vars.items():
                            if pk in plist:
                                if isinstance(pvar, tk.BooleanVar): pvar.set(bool(plist[pk]))
                                else: pvar.set(str(plist[pk]))
                        # Log success for key fine-tune variables
                        self.log(f"[Restore] 사진대장 설정 복구: 비율({self.photo_width_pct_var.get()}%), 추가({self.photo_width_pixel_adj_var.get()}px), 이동({self.photo_shift_x_var.get()}px)")
                        if 'selected_files' in plist:
                            self.photo_selected_files = plist['selected_files']
                            if hasattr(self, 'photo_listbox'):
                                self.photo_listbox.delete(0, tk.END)
                                for f_path in self.photo_selected_files:
                                    self.photo_listbox.insert(tk.END, f_path)

                    self.config.update(saved_data)
                self.log("[SUCCESS] 사용자 저장 설정을 모두 불러왔습니다.")
                
                # Restore Window Geometry/State
                if 'WINDOW_GEOMETRY' in saved_data:
                    self.root.geometry(saved_data['WINDOW_GEOMETRY'])
                if 'WINDOW_STATE' in saved_data:
                    try: self.root.state(saved_data['WINDOW_STATE'])
                    except: pass
                if 'ACTIVE_TAB' in saved_data:
                    self.root.after(200, lambda: self.mode_notebook.select(saved_data['ACTIVE_TAB']))
            except Exception as e:
                print(f"WARNING: 사용자 설정 불러오기 실패: {e}")

        # PAUT 마이그레이션 (UT_* -> PAUT_*)
        for k in list(self.config.keys()):
            if k.startswith("UT_"):
                new_k = k.replace("UT_", "PAUT_")
                if new_k not in self.config:
                    self.config[new_k] = self.config[k]
        
        # [NEW] Load Photo Log specific config
        self.load_photo_log_config()

    def capture_ui_state(self):
        """UI 요소들의 현재 상태(컬럼 너비, 분할선 위치 등)를 config에 반영"""
        try:
            trees = [
                ('PMI', getattr(self, 'preview_tree', None)), 
                ('RT', getattr(self, 'rt_preview_tree', None)), 
                ('PT', getattr(self, 'pt_preview_tree', None)), 
                ('PAUT', getattr(self, 'paut_preview_tree', None)),
                ('KOGAS', getattr(self, 'kogas_preview_tree', None))
            ]
            for mode, tree in trees:
                if tree:
                    try:
                        self.config[f"{mode}_COL_WIDTHS"] = {col: tree.column(col, "width") for col in tree["columns"]}
                    except: pass

            # 2. 패널 분할선(Sash) 비율 캡처
            paned_windows = [
                ('PMI', getattr(self, 'pmi_paned', None)),
                ('RT', getattr(self, 'rt_paned', None)),
                ('PT', getattr(self, 'pt_paned', None)),
                ('PAUT', getattr(self, 'paut_paned', None))
            ]
            for mode, pw in paned_windows:
                if pw:
                    try:
                        total_w = pw.winfo_width()
                        if total_w > 100:
                            self.config[f'{mode}_SASH_RATIO'] = pw.sash_coord(0)[0] / total_w
                    except: pass

            # 3. 윈도우 및 활성 탭 상태 캡처
            try:
                self.config['WINDOW_GEOMETRY'] = self.root.geometry()
                self.config['WINDOW_STATE'] = self.root.state()
                if hasattr(self, 'mode_notebook'):
                    self.config['ACTIVE_TAB'] = self.mode_notebook.index("current")
                
                # 4. 파일 경로 관성 저장
                self.config['PMI_TARGET_PATH'] = self.target_file_path.get()
                self.config['PMI_TEMPLATE_PATH'] = self.template_file_path.get()
                self.config['RT_TARGET_PATH'] = self.rt_target_file_path.get()
                self.config['RT_TEMPLATE_PATH'] = self.rt_template_file_path.get()
                self.config['PT_TARGET_PATH'] = self.pt_target_file_path.get()
                self.config['PT_TEMPLATE_PATH'] = self.pt_template_file_path.get()
                self.config['PAUT_TARGET_PATH'] = self.paut_target_file_path.get()
                self.config['PAUT_TEMPLATE_PATH'] = self.paut_template_file_path.get()
                self.config['KOGAS_TARGET_PATH'] = self.kogas_target_file_path.get()
                self.config['KOGAS_TEMPLATE_PATH'] = self.kogas_template_file_path.get()
                
                # 5. Photo Log Settings Capture
                self.config['PHOTO_LOG_SETTINGS'] = {
                    'orderer': self.photo_orderer.get(), 'inspect_date': self.photo_inspect_date.get(),
                    'inspect_type': self.photo_inspect_type.get(), 'report_title': self.photo_report_title.get(),
                    'report_no': self.photo_report_no.get(),
                    'cols_per_row': self.photo_cols_per_row.get(), 'keep_aspect': self.photo_keep_aspect.get(),
                    'output_name': self.photo_output_name.get(), 'logo_path': self.photo_logo_path.get(),
                    'logo_width': self.photo_logo_width_var.get(),
                    'logo_x': self.photo_logo_x_var.get(), 'logo_y': self.photo_logo_y_var.get(),
                    'cell_width': self.photo_cell_width_var.get(), 'cell_height': self.photo_cell_height_var.get(),
                    'm_top': self.photo_margin_top_var.get(), 'm_bottom': self.photo_margin_bottom_var.get(),
                    'm_left': self.photo_margin_left_var.get(), 'm_right': self.photo_margin_right_var.get(),
                    'print_scale': self.photo_print_scale_var.get(), 'desc_height': self.photo_desc_height_var.get(),
                    'photo_align': self.photo_align_var.get(), 'fit_width': self.photo_fit_width_var.get(),
                    'auto_rotate': self.photo_auto_rotate_var.get(), 'width_pct': self.photo_width_pct_var.get(),
                    'pixel_adj': self.photo_width_pixel_adj_var.get(), 'shift_x': self.photo_shift_x_var.get(),
                    'shift_y': self.photo_shift_y_var.get(),
                    'selected_files': self.photo_selected_files,
                    'last_save_dir': getattr(self, 'last_photo_save_dir', "")
                }
            except Exception as e:
                self.log(f"[ERROR] UI 상태 캡처 실패: {e}")
        except Exception as e:
            self.log(f"[ERROR] 전체 상태 캡처 실패: {e}")

    def save_photo_log_config(self):
        """Save Photo Log specific layout settings to a separate JSON."""
        config_path = os.path.join(os.getcwd(), "photolog_config.json")
        self.capture_ui_state()
        plist = self.config.get('PHOTO_LOG_SETTINGS', {})
        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(plist, f, indent=4, ensure_ascii=False)
            self.log("💾 사진대장 전용 설정이 저장되었습니다.")
            messagebox.showinfo("저장 완료", "사진대장 레이아웃 설정이 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패: {e}")

    def load_photo_log_config(self):
        """Load Photo Log specific layout settings."""
        config_path = os.path.join(os.getcwd(), "photolog_config.json")
        if not os.path.exists(config_path): return
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                plist = json.load(f)
            
            # Update UI vars (using same logic as in load_settings)
            photo_vars = {
                'orderer': self.photo_orderer, 'inspect_date': self.photo_inspect_date,
                'inspect_type': self.photo_inspect_type, 'report_title': self.photo_report_title,
                'report_no': self.photo_report_no,
                'cols_per_row': self.photo_cols_per_row, 'keep_aspect': self.photo_keep_aspect,
                'output_name': self.photo_output_name, 'logo_path': self.photo_logo_path,
                'logo_width': self.photo_logo_width_var,
                'logo_x': self.photo_logo_x_var, 'logo_y': self.photo_logo_y_var,
                'cell_width': self.photo_cell_width_var, 'cell_height': self.photo_cell_height_var,
                'm_top': self.photo_margin_top_var, 'm_bottom': self.photo_margin_bottom_var,
                'm_left': self.photo_margin_left_var, 'm_right': self.photo_margin_right_var,
                'print_scale': self.photo_print_scale_var, 'desc_height': self.photo_desc_height_var,
                'photo_align': self.photo_align_var, 'fit_width': self.photo_fit_width_var,
                'auto_rotate': self.photo_auto_rotate_var, 'width_pct': self.photo_width_pct_var,
                'pixel_adj': self.photo_width_pixel_adj_var, 'shift_x': self.photo_shift_x_var,
                'shift_y': self.photo_shift_y_var
            }
            for pk, pvar in photo_vars.items():
                if pk in plist:
                    if isinstance(pvar, tk.BooleanVar): pvar.set(bool(plist[pk]))
                    else: pvar.set(str(plist[pk]))
            
            if 'selected_files' in plist:
                self.photo_selected_files = plist['selected_files']
                if hasattr(self, 'photo_listbox'):
                    self.photo_listbox.delete(0, tk.END)
                    for f_path in self.photo_selected_files:
                        self.photo_listbox.insert(tk.END, f_path)
            
            self.log("📂 사진대장 전용 설정을 불러왔습니다.")
        except Exception as e:
            print(f"Error loading photolog config: {e}")

    def load_template_specific_config(self, template_path, mode="RT"):
        """템플릿 파일과 연동된 JSON 설정 파일을 로드하여 UI에 반영"""
        if not template_path or not os.path.exists(template_path): return
        
        config_path = os.path.splitext(template_path)[0] + ".json"
        if not os.path.exists(config_path): return
        
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                t_config = json.load(f)
            
            self.log(f"📂 [Auto-Load] '{os.path.basename(template_path)}' 전용 설정을 불러왔습니다.")
            
            # Update self.config and UI variables
            for key, value in t_config.items():
                self.config[key] = value
                if key in self.setting_vars:
                    self.setting_vars[key].set(str(value))
            
            # Mode-specific UI update (Force refresh logic if needed)
            if mode in ["RT", "KOGAS"]:
                self._update_rt_preview_columns(mode)
            elif mode == "PT":
                try: self._update_pt_preview_columns()
                except: pass
            elif mode == "PAUT":
                try: self._update_paut_target_fsh()
                except: pass
            elif mode == "PMI":
                try: self._update_pmi_filter_ui()
                except: pass

        except Exception as e:
            self.log(f"[ERROR] 템플릿 전용 설정 로드 실패: {e}")

    def save_template_specific_config(self, mode="RT"):
        """현재 설정을 선택된 템플릿 전용 JSON 파일로 저장"""
        if mode == "RT": template_path = self.rt_template_file_path.get()
        elif mode == "KOGAS": template_path = self.kogas_template_file_path.get()
        elif mode == "PT": template_path = self.pt_template_file_path.get()
        elif mode == "PAUT": template_path = self.paut_template_file_path.get()
        else: template_path = self.template_file_path.get()
        
        if not template_path or not os.path.exists(template_path):
            messagebox.showwarning("경고", "먼저 템플릿 파일을 선택해주세요.")
            return
            
        config_path = os.path.splitext(template_path)[0] + ".json"
        self.capture_ui_state() # Ensure current UI is in self.config
        
        # Determine relevant keys to save for this template
        # 1. Logos & Print settings (Template-dependent)
        keys_to_save = []
        for ctx in ["COVER", "DATA"]:
            # Correct Prefix mapping: SITCO_RT_COVER, etc.
            # Actually, mode is like "RT", ctx is "COVER" -> prefix is "RT_COVER"
            # In UI: self._create_setting_grid(t_cover, "RT_COVER")
            # Inside grid: keys are SITCO_RT_COVER_X etc.
            for prefix in ["SITCO", "SEOUL", "FOOTER", "FOOTER_PT"]:
                for suffix in ["_PATH", "_ANCHOR", "_W", "_H", "_X", "_Y"]:
                    keys_to_save.append(f"{prefix}_{mode}_{ctx}{suffix}")
            for suffix in ["_TOP", "_BOTTOM", "_LEFT", "_RIGHT"]:
                keys_to_save.append(f"MARGIN_{mode}_{ctx}{suffix}")
            keys_to_save.append(f"PRINT_SCALE_{mode}_{ctx}")
            keys_to_save.append(f"PRINT_AREA_{mode}_{ctx}")
            
        # 2. Mode-specific boundaries and column mappings
        if mode == "RT":
            keys_to_save += ["RT_START_ROW", "RT_END_ROW"]
            keys_to_save += [k for k in self.config.keys() if k.startswith("RT_COL_") or k.startswith("RT_NAME_")]
        elif mode == "KOGAS":
            keys_to_save += ["KOGAS_START_ROW", "KOGAS_DATA_END_ROW", "RT_KOGAS_D_START_COL"]
            keys_to_save += [k for k in self.config.keys() if k.startswith("KOGAS_COL_") or k.startswith("KOGAS_NAME_")]
        elif mode == "PT":
            keys_to_save += ["PT_START_ROW", "PT_END_ROW"]
            keys_to_save += [k for k in self.config.keys() if k.startswith("PT_COL_") or k.startswith("PT_NAME_")]
        elif mode == "PAUT":
            keys_to_save += ["PAUT_START_ROW", "PAUT_END_ROW"]
        elif mode == "PMI":
            keys_to_save += ["PMI_START_ROW", "PMI_DATA_END_ROW", "PMI_PRINT_END_ROW"]
            keys_to_save += [k for k in self.config.keys() if k.startswith("PMI_COL_") or k.startswith("PMI_NAME_")]

        # Build template-specific dict
        t_data = {k: self.config[k] for k in keys_to_save if k in self.config}
        
        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(t_data, f, indent=4, ensure_ascii=False)
            self.log(f"💾 [SUCCESS] '{os.path.basename(template_path)}' 전용 설정이 저장되었습니다.")
            messagebox.showinfo("저장 완료", f"템플릿 전용 설정이 저장되었습니다.\n{os.path.basename(config_path)}")
        except Exception as e:
            messagebox.showerror("저장 실패", f"설정 저장 중 오류 발생: {e}")

    def save_settings(self):
        """현재 설정을 파일(JSON)에 저장"""
        self.capture_ui_state()
        try:
            if hasattr(self, 'setting_vars'):
                for key, var in self.setting_vars.items():
                    val = var.get()
                    try:
                        # [FIX] Handle various key types correctly for persistence
                        if "AREA" in key or "_PATH" in key:
                            self.config[key] = str(val).strip()
                        elif any(x in key for x in ['_X', '_Y', '_W', '_H', 'MARGIN', 'HEIGHT', 'WIDTH']):
                            try: self.config[key] = float(val) if str(val).strip() else 0.0
                            except: self.config[key] = 0.0
                        elif 'SCALE' in key or key.endswith(('_ROW', '_COL', '_IDX')):
                            # Column indexes and row numbers should be handled as strings if they are like 'A', 
                            # but for safety in logic we often use col_to_num. 
                            # Here we store them as strings to preserve 'A', 'B' etc.
                            self.config[key] = str(val).strip()
                        else:
                            self.config[key] = str(val)
                    except Exception as e:
                        self.config[key] = str(val)

            # [NEW] 즉시 파일로 기록하여 리셋 방지
            config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'config', 'logo_settings_unified.json')
            os.makedirs(os.path.dirname(config_path), exist_ok=True)
            try:
                with open(config_path, 'w', encoding='utf-8') as f:
                    json.dump(self.config, f, indent=4, ensure_ascii=False)
            except: pass

            # [NEW] 필터 설정 저장 (PMI)
            if hasattr(self, 'element_filters'):
                filter_data = []
                for f_item in self.element_filters:
                    filter_data.append({
                        'key': f_item['key'].get(),
                        'min': f_item['min'].get(),
                        'max': f_item['max'].get()
                    })
                self.config["PMI_FILTERS"] = filter_data

            # [NEW] 탭별 전용 행 설정 강제 저장
            self.config["PMI_START_ROW"] = self.config.get("START_ROW", 19)
            self.config["PMI_DATA_END_ROW"] = self.config.get("DATA_END_ROW", 45)
            self.config["PMI_PRINT_END_ROW"] = self.config.get("PRINT_END_ROW", 47)

            # [NEW] Window State Capture
            self.config['WINDOW_GEOMETRY'] = self.root.geometry()
            self.config['WINDOW_STATE'] = self.root.state()
            try: self.config['ACTIVE_TAB'] = self.mode_notebook.index("current")
            except: pass
            # [FIX] Ensure the config directory exists (Prevents Errno 2)
            os.makedirs(CONFIG_DIR, exist_ok=True)

            with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=4, default=str)
            self.log("[SUCCESS] 모든 설정이 파일에 안전하게 저장되었습니다.")
            return True, "성공"
        except Exception as e:
            import traceback
            err_msg = traceback.format_exc()
            self.log(f"[ERROR] 설정 저장 실패 상세:\n{err_msg}")
            return False, str(e)

    def manual_save_settings(self):
        """사용자가 수동으로 버튼을 눌러 저장할 때 알림 표시"""
        success, msg = self.save_settings()
        if success:
            messagebox.showinfo("성공", "현재 설정이 안전하게 저장되었습니다.")
        else:
            messagebox.showerror("오류", f"설정 저장 중 문제가 발생했습니다.\n\n원인: {msg}")

    def on_closing(self):
        """애플리케이션 종료 시 최종 상태 저장"""
        try:
            self.capture_ui_state()
            self.save_settings()
        except: pass
        self.root.destroy()

    def _apply_sash_ratio(self, mode):
        """Helper to apply saved sash ratio with a slight delay to ensure UI mapping."""
        pw = getattr(self, f"{mode.lower()}_paned", None)
        ratio = self.config.get(f"{mode}_SASH_RATIO")
        if pw and ratio is not None:
            def apply():
                total_w = pw.winfo_width()
                if total_w > 100:
                    pw.sash_place(0, int(total_w * ratio), 0)
                else:
                    self.root.after(100, apply)
            self.root.after(100, apply)

    def evaluate_paut_flaw(self, t, h, l, depth, flaw_nature, mode="ECA"):
        """
        PAUT Acceptance Criteria Logic
        - Standard (ASME B31.1): Fatal flaws (Crack, LOF, IP) are rejected immediately.
        - ECA (Fracture Mechanics): Only Crack is fatal. LOF/IP are evaluated by size.
        """
        try:
            t_val = float(t)
            h_val = float(h)
            l_val = float(l)
            d_val = float(depth)
        except (ValueError, TypeError):
            return "Error (Invalid Dimension)", "Unknown"

        # 0. Automatic Location Determination
        s_top = d_val
        s_bottom = t_val - (d_val + h_val)
        s = min(s_top, s_bottom)
        s_limit = 0.4 * (h_val / 2)
        
        loc = "Surface" if s <= s_limit else "Subsurface"

        # 1. Immediate Rejection (Crack)
        nature_str = str(flaw_nature).strip().lower()
        if 'crack' in nature_str or '균열' in nature_str:
            return "Reject (Crack)", loc
            
        # [NEW] Mode-based Rejection for LOF (LF) and IP
        if mode == "Standard":
            unacceptable_types = ['lof', 'lack of fusion', 'ip', 'incomplete penetration', 'lf']
            if any(x in nature_str for x in unacceptable_types):
                return f"Reject ({flaw_nature})", loc
        
        if l_val <= 0 or h_val <= 0 or t_val <= 0:
            return "Error (Zero/Negative Value)", loc

        # 1.1 Special Rules for 6mm <= t < 13mm
        if 6 <= t_val < 13:
            if l_val > 6.4: return f"Reject (L: {l_val} > 6.4mm)", loc
            if t_val < 10: h_surf_max, h_sub_max = 0.95, 0.96
            elif t_val < 12: h_surf_max, h_sub_max = 1.04, 1.04
            else: h_surf_max, h_sub_max = 1.13, 1.14
                
            limit = h_surf_max if loc == "Surface" else h_sub_max
            if h_val > limit: return f"Reject ({loc} h: {h_val} > {limit}mm)", loc
            return "Accept", loc

        # 1.2 Special Rules for 13mm <= t < 25.4mm
        if 13 <= t_val < 25.4:
            if l_val > 6.4: return f"Reject (L: {l_val} > 6.4mm)", loc
            actual_h_t = h_val / t_val
            allowed_h_t = 0.087 if loc == "Surface" else 0.143 
            if actual_h_t > allowed_h_t: return f"Reject ({loc} h/t: {actual_h_t:.3f} > {allowed_h_t:.3f})", loc
            return "Accept", loc
                
        # 2. Aspect Ratio (a/l) logic for t >= 25.4mm
        a_val = h_val if loc == "Surface" else h_val / 2
        aspect_ratio_a_l = a_val / l_val
        
        master_table = [
            (0.00, 0.031, 0.034), (0.05, 0.033, 0.038), (0.10, 0.036, 0.043),
            (0.15, 0.041, 0.054), (0.20, 0.047, 0.066), (0.25, 0.055, 0.078),
            (0.30, 0.064, 0.090), (0.35, 0.074, 0.103), (0.40, 0.083, 0.116),
            (0.45, 0.085, 0.129), (0.50, 0.087, 0.143)
        ]
        
        allowed_a_t = 0
        for ar_limit, surf_a_t, sub_a_t in master_table:
            if aspect_ratio_a_l <= ar_limit:
                allowed_a_t = surf_a_t if loc == 'Surface' else sub_a_t
                break
        if allowed_a_t == 0: allowed_a_t = master_table[-1][1] if loc == 'Surface' else master_table[-1][2]

        actual_a_t = a_val / t_val
        if actual_a_t <= allowed_a_t: return "Accept", loc
        return f"Reject ({loc} a/t: {actual_a_t:.3f} > {allowed_a_t:.3f})", loc

    def _on_main_tab_changed(self, event):
        try:
            tab_text = self.mode_notebook.tab(self.mode_notebook.select(), "text")
            if "PMI" in tab_text: self.current_mode = "PMI"
            elif "RT" in tab_text:
                # RT 탭 진입 시 현재 설정된 서브 모드(표준/가스공사) 확인
                self.current_mode = self.rt_sub_mode.get()
            elif "PT" in tab_text: self.current_mode = "PT"
            elif "PAUT" in tab_text: self.current_mode = "PAUT"
            elif "사진" in tab_text: self.current_mode = "PHOTO"
            
            # [NEW] 모드 변경 시 해당 모드의 설정을 다시 로드
            if self.current_mode in ["RT", "KOGAS"]:
                t_path = self.rt_template_file_path.get() if self.current_mode == "RT" else self.kogas_template_file_path.get()
                self.load_template_specific_config(t_path, self.current_mode)
        except: pass

    def create_widgets(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TFrame", background="#f9fafb")
        style.configure("TLabel", background="#f9fafb", font=("Malgun Gothic", 10))
        style.configure("TLabelframe", background="#f9fafb", font=("Malgun Gothic", 10, "bold"))
        style.configure("TLabelframe.Label", background="#f9fafb", font=("Malgun Gothic", 10, "bold"))
        style.configure("Action.TButton", font=("Malgun Gothic", 11, "bold"), padding=10)
        
        # [NEW] 탭 스타일 설정
        style.configure("Main.TNotebook.Tab", font=("Malgun Gothic", 11, "bold"), padding=[10, 3])
        
        style.map("TEntry", 
                  selectbackground=[('focus', '#3b82f6'), ('!focus', '#3b82f6')],
                  selectforeground=[('focus', 'white'), ('!focus', 'white')])

        # --- Main Scrollable Container ---
        self.canvas = tk.Canvas(self.root, background="#f9fafb", highlightthickness=0, yscrollincrement=40)
        self.scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, background="#f9fafb")

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self._safe_update_scrollregion()
        )

        self.canvas_frame_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        def _on_canvas_configure(event):
            # [ROOT CAUSE FIX] Always force width, handle None bbox gracefully
            self.canvas.itemconfigure(self.canvas_frame_window, width=event.width)
            self._safe_update_scrollregion()

        self.canvas.bind("<Configure>", _on_canvas_configure)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        self.root.option_add('*selectBackground', '#3b82f6')
        self.root.option_add('*selectForeground', 'white')

        # Treeview Style
        style.configure("Treeview", rowheight=28, font=("Malgun Gothic", 10))
        style.map("Treeview", background=[('selected', '#3b82f6')], foreground=[('selected', 'white')])
        style.configure("Treeview.Heading", font=("Malgun Gothic", 10, "bold"))

        self._create_entry_context_menu()
        self._create_tree_context_menu()

        # Mouse Wheel Support
        def _on_mousewheel(event):
            try:
                focused = self.root.focus_get()
                curr = focused
                while curr:
                    try:
                        c_name = curr.winfo_class()
                        if c_name in ('Treeview', 'Text', 'Listbox'):
                            curr.yview_scroll(int(-1*(event.delta/120)*4.0), "units")
                            return "break"
                    except: pass
                    curr = curr.master
            except Exception: pass
            self.canvas.yview_scroll(int(-1*(event.delta/120)*2.0), "units")
            return "break"
        self.root.bind_all("<MouseWheel>", _on_mousewheel, add="+")

        def _on_root_click(event):
            try:
                if "treeview" in str(event.widget).lower(): return
                w_class = event.widget.winfo_class()
                if w_class in ('Frame', 'TFrame', 'Label', 'TLabel', 'Canvas', 'Labelframe', 'TLabelframe'):
                    self.root.focus_set()
            except: pass
        self.root.bind_all("<Button-1>", _on_root_click, add="+")

        # [ROOT CAUSE FIX] Place Notebook DIRECTLY in root, NOT in Canvas scrollable_frame.
        # Canvas windows do not reliably propagate width to children.
        # Since PMI uses splitscreen (no scrolling needed), direct packing guarantees full width.
        
        # Hide Canvas & Scrollbar (no longer needed for layout)
        self.canvas.pack_forget()
        self.scrollbar.pack_forget()
        
        # ===== PACK ORDER: bottom-most items first =====
        
        # [UX] ESC Key Binding: Unfocus only (preserving content)
        self.root.bind_class("TEntry", "<Escape>", self._on_entry_esc)
        self.root.bind_class("Entry", "<Escape>", self._on_entry_esc)
        
        # 1. Status bar (very bottom)
        self.status_bar = tk.Frame(self.root, background="#e2e8f0", height=22)
        self.status_bar.pack(fill='x', side='bottom')
        self.status_bar.pack_propagate(False)
        tk.Label(self.status_bar, text="준비됨", font=("Malgun Gothic", 8), background="#e2e8f0", foreground="#64748b").pack(side='left', padx=10)
        tk.Label(self.status_bar, text=f"Build Version: v{APP_VERSION}", font=("Arial", 8), background="#e2e8f0", foreground="#94a3b8").pack(side='right', padx=10)

        # 2. Progress bar + Log (above status bar, compact)
        bottom_frame = tk.Frame(self.root, background="#f9fafb")
        bottom_frame.pack(fill='x', side='bottom', padx=5)

        self.progress = ttk.Progressbar(bottom_frame, orient="horizontal", mode="determinate")
        self.progress.pack(fill='x', pady=(5, 2))

        log_frame = tk.Frame(bottom_frame, background="#f9fafb")
        log_frame.pack(fill='x')

        self.status_log = tk.Text(log_frame, height=4, font=("Consolas", 8), state='disabled', background="#1e1e2e", foreground="#10b981", padx=5, pady=3)
        vsb = ttk.Scrollbar(log_frame, orient="vertical", command=self.status_log.yview)
        self.status_log.configure(yscrollcommand=vsb.set)
        self.status_log.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        
        # 3. THEN pack Notebook (fills all remaining space)
        self.mode_notebook = ttk.Notebook(self.root, style="Main.TNotebook")
        self.mode_notebook.pack(fill='both', expand=True, padx=5, pady=(5, 0))
        
        # [NEW] Track mode change
        self.mode_notebook.bind("<<NotebookTabChanged>>", self._on_main_tab_changed)
        
        self.pmi_mode_frame = tk.Frame(self.mode_notebook, background="#f9fafb")
        self.rt_mode_frame = tk.Frame(self.mode_notebook, background="#f9fafb")
        self.pt_mode_frame = tk.Frame(self.mode_notebook, background="#f9fafb")
        self.paut_mode_frame = tk.Frame(self.mode_notebook, background="#f9fafb")
        self.photo_mode_frame = tk.Frame(self.mode_notebook, background="#f9fafb")

        self.mode_notebook.add(self.pmi_mode_frame, text=" 🔬 PMI (OES) ")
        self.mode_notebook.add(self.rt_mode_frame, text=" 🔬 RT (Standard) ")
        

        
        self.pt_mode_frame = tk.Frame(self.mode_notebook, background="#f9fafb")
        self.mode_notebook.add(self.pt_mode_frame, text=" 🔬 Penetrant (PT) ")
        self.mode_notebook.add(self.paut_mode_frame, text=" 🔬 Phased Array (PAUT) ")
        self.mode_notebook.add(self.photo_mode_frame, text=" 📸 사진대장 (Photo Log) ")

        # Setup each mode (One time only)
        self._setup_pmi_ui(self.pmi_mode_frame)
        self._setup_rt_ui(self.rt_mode_frame) # [RE-INTEGRATED]
        self._setup_pt_ui(self.pt_mode_frame)
        self._setup_paut_ui(self.paut_mode_frame)
        self._setup_photo_log_ui(self.photo_mode_frame)

    def _create_scrollable_sidebar(self, parent):
        """Creates a scrollable canvas/scrollbar container for sidebars."""
        canvas = tk.Canvas(parent, background="#f9fafb", highlightthickness=0, borderwidth=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, background="#f9fafb", padx=10, pady=0, highlightthickness=0, borderwidth=0)

        def _update_scrollregion(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Auto-hide scrollbar if content fits
            if canvas.bbox("all")[3] <= canvas.winfo_height():
                scrollbar.pack_forget()
            else:
                scrollbar.pack(side="right", fill="y")
        
        scrollable_frame.bind("<Configure>", _update_scrollregion)

        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        def _on_canvas_configure(e):
            canvas.itemconfig(canvas_window, width=e.width)
        canvas.bind("<Configure>", _on_canvas_configure)
        
        canvas.configure(yscrollcommand=scrollbar.set)
        
        def _on_mousewheel(event):
            # Only allow scrolling if content actually exceeds the visible area
            if canvas.bbox("all")[3] > canvas.winfo_height():
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            return "break"
        
        canvas.bind('<Enter>', lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
        canvas.bind('<Leave>', lambda e: canvas.unbind_all("<MouseWheel>"))

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        return scrollable_frame

    def _setup_pmi_ui(self, parent):
        # [FORCE] Ensure parent (pmi_mode_frame) allows expansion
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # --- Dual Pane Layout (PanedWindow) ---
        self.pmi_paned = tk.PanedWindow(container, orient='horizontal', background="#d1d5db", 
                            sashwidth=6, sashpad=0, sashrelief='raised', borderwidth=0)
        self.pmi_paned.pack(fill='both', expand=True)

        # [LEFT] Filter/Settings Sidebar
        left_container = tk.Frame(self.pmi_paned, background="#f9fafb", highlightthickness=0, borderwidth=0)
        self.pmi_paned.add(left_container, width=425)

        # Scrollable area (Full size)
        left_pane = self._create_scrollable_sidebar(left_container)

        # FIXED FLOATING Header (Lifts over scrollable area)
        header_frame = tk.Frame(left_container, background="#f9fafb", highlightthickness=0, borderwidth=0)
        header_frame.place(x=0, y=0, relwidth=1, height=40)
        
        inner_header = tk.Frame(header_frame, background="#f9fafb", padx=20)
        inner_header.pack(fill='both', expand=True, pady=(5, 0))
        
        tk.Label(inner_header, text="🔬 PMI 관리", font=("Malgun Gothic", 15, "bold"), 
                 background="#f9fafb", foreground="#1e3a8a").pack(side='left')
        tk.Label(inner_header, text=f"v{APP_VERSION}", font=("Arial", 8), 
                 background="#f1f5f9", foreground="#64748b", padx=5, pady=0).pack(side='left', padx=10)
        
        # Bottom separator
        tk.Frame(header_frame, height=1, background="#e5e7eb").pack(side='bottom', fill='x')
        header_frame.lift() # Ensure it's on top

        # Spacer in scrollable area so first group starts below floating header
        tk.Frame(left_pane, height=40, background="#f9fafb").pack(fill='x')

        # 1. File Selection Group
        file_container = ttk.LabelFrame(left_pane, text=" 데이터 및 양식 (Data) ", padding=10)
        file_container.pack(fill='x', pady=(0, 10))

        def _add_very_compact_row(parent_f, label, var, row, is_dir=False, types=None):
            parent_f.columnconfigure(1, weight=1)
            ttk.Label(parent_f, text=label, font=("Malgun Gothic", 8)).grid(row=row, column=0, sticky='e', padx=2, pady=2)
            # Use width=1 for entry to allow proactive shrinking
            ttk.Entry(parent_f, textvariable=var, font=("Arial", 9), width=1, exportselection=False).grid(row=row, column=1, padx=2, pady=2, sticky='ew')
            cmd = (lambda : self._browse_dir(var)) if is_dir else (lambda : self._browse_file(var, types))
            ttk.Button(parent_f, text="...", width=3, command=cmd).grid(row=row, column=2, padx=2, pady=2)

        _add_very_compact_row(file_container, "기본 로고 폴더:", self.logo_folder_path, 0, is_dir=True)
        _add_very_compact_row(file_container, "데이터:", self.target_file_path, 1, types=[("Excel Source", "*.xls;*.xlsx;*.xlsm")])
        _add_very_compact_row(file_container, "양식:", self.template_file_path, 2, types=[("Excel Template", "*.xlsx;*.xlsm")])
        
        # [NEW] Template Config Save Button
        btn_frame = tk.Frame(file_container, background="#f0f0f0")
        btn_frame.grid(row=3, column=1, sticky='ew', pady=2)
        ttk.Button(btn_frame, text="💾 현재 설정을 이 양식 전용으로 저장", command=lambda: self.save_template_specific_config("PMI")).pack(side='left', padx=2)

        # 2. Configuration Notebook (Cover, Data, Rows - NO Preview tab here)
        config_frame = ttk.LabelFrame(left_pane, text=" 설정 (Config) ", padding=2)
        config_frame.pack(fill='both', expand=True, pady=(0, 10))

        self.pmi_tab_notebook = ttk.Notebook(config_frame)
        self.pmi_tab_notebook.pack(fill='both', expand=True)

        tab_cover = ttk.Frame(self.pmi_tab_notebook, padding=5)
        tab_data = ttk.Frame(self.pmi_tab_notebook, padding=5)
        tab_rows = ttk.Frame(self.pmi_tab_notebook, padding=5)
        tab_filter = ttk.Frame(self.pmi_tab_notebook, padding=10)
        tab_cols = ttk.Frame(self.pmi_tab_notebook, padding=5)
        
        # [CRITICAL] Allow children to expand within tabs
        for t in [tab_cover, tab_data, tab_rows, tab_filter, tab_cols]: t.columnconfigure(0, weight=1)

        self.pmi_tab_notebook.add(tab_cover, text="갑지")
        self.pmi_tab_notebook.add(tab_data, text="을지")
        self.pmi_tab_notebook.add(tab_rows, text="행 설정")
        self.pmi_tab_notebook.add(tab_cols, text="컬럼 설정")
        self.pmi_tab_notebook.add(tab_filter, text="필터/옵션")

        # [FIX] Do NOT reset setting_vars here to prevent data loss when switching tabs
        # [ALIGNED] Mode-specific context for logo grid
        self._create_gapji_meta_ui(tab_cover, use_pack=False)
        self.pmi_tab_notebook.bind("<<NotebookTabChanged>>", self._update_gapji_preview_current)
        next_row_cover = self._create_setting_grid(tab_cover, "PMI_COVER")
        next_row_data = self._create_setting_grid(tab_data, "PMI_DATA")
        self._create_margin_settings(tab_cover, "PMI_COVER", use_pack=False)
        self._create_margin_settings(tab_data, "PMI_DATA", use_pack=False)
        self._create_row_settings(tab_rows, mode="PMI")
        
        pmi_items = [
            ("No:", "PMI_COL_NO", 13, "PMI_NAME_NO", "No", "No"),
            ("Drawing No:", "PMI_COL_DWG", 1, "PMI_NAME_DWG", "Drawing No.", "Dwg"),
            ("Joint No:", "PMI_COL_JOINT", 6, "PMI_NAME_JOINT", "Joint No.", "Joint"),
            ("Location:", "PMI_COL_LOC", 7, "PMI_NAME_LOC", "Location", "Loc"),
            ("Ni:", "PMI_COL_NI", 8, "PMI_NAME_NI", "Ni", "Ni"),
            ("Cr:", "PMI_COL_CR", 9, "PMI_NAME_CR", "Cr", "Cr"),
            ("Mo:", "PMI_COL_MO", 10, "PMI_NAME_MO", "Mo", "Mo"),
            ("판정 결과:", "PMI_COL_RES", 13, "PMI_NAME_RES", "Result", "Result")
        ]
        self._create_column_mapping_ui(tab_cols, "PMI", pmi_items)
        
        # Spacer for tabs
        tk.Frame(tab_rows, background="#f9fafb").pack(fill='x')

        # 3. Quick Options Section (Moved to Settings Notebook)
        # Filter Tag Area (Self-updating)
        filter_box = tk.Frame(tab_filter, background="#f9fafb")
        filter_box.pack(fill='x', pady=2)
        ttk.Label(filter_box, text="🔍 성분 필터:", background="#f9fafb", font=("Malgun Gothic", 9, "bold")).pack(side='left', padx=(0, 5))
        
        def add_filter():
            self.element_filters.append({'key': tk.StringVar(value=''), 'min': tk.StringVar(value=''), 'max': tk.StringVar(value='')})
            self._update_pmi_filter_ui()
        ttk.Button(filter_box, text="+", width=2, command=add_filter).pack(side='right')

        self.filter_container = tk.Frame(filter_box, background="#f9fafb")
        self.filter_container.pack(side='left', fill='x', expand=True)

        # Sequence Filter & Auto Check
        tk.Frame(tab_filter, height=1, background="#e5e7eb").pack(fill='x', pady=10)
        tk.Checkbutton(tab_filter, text="✅ 재질 자동 판정", variable=self.auto_verify, background="#f9fafb", font=("Malgun Gothic", 8)).pack(anchor='w', pady=(0, 5))
        seq_row = tk.Frame(tab_filter, background="#f9fafb")
        seq_row.pack(fill='x', pady=2)
        ttk.Label(seq_row, text="📊 특정순번:", background="#f9fafb", font=("Malgun Gothic", 8)).pack(side='left')
        ttk.Entry(seq_row, textvariable=self.sequence_filter, width=15, font=("Arial", 9), exportselection=False).pack(side='left', padx=5, fill='x', expand=True)

        action_bar = tk.Frame(left_pane, background="#f9fafb")
        action_bar.pack(fill='x', pady=5)
        ttk.Button(action_bar, text=" ✨ 생성 시작 ", style="Action.TButton", command=self.run_process).pack(fill='x', pady=(0, 5))
        ttk.Button(action_bar, text=" 📝 데이터 추출 ", command=self.extract_only).pack(fill='x')

        self._update_pmi_filter_ui()

        # [RIGHT] Multi-Preview Pane (Data & Gapji)
        right_container = tk.Frame(self.pmi_paned, background="#f3f4f6")
        self.pmi_paned.add(right_container, stretch="always")
        
        if not hasattr(self, 'preview_notebooks'): self.preview_notebooks = {}
        nb = ttk.Notebook(right_container)
        nb.pack(fill='both', expand=True)
        self.preview_notebooks["PMI"] = nb
        
        tab_data = tk.Frame(nb, background="#ffffff")
        tab_gapji = tk.Frame(nb, background="#ffffff")
        nb.add(tab_data, text=" 🔬 데이터 미리보기 ")
        nb.add(tab_gapji, text=" 📄 갑지 미리보기 ")
        
        self._create_preview_ui(tab_data)
        self._create_gapji_preview_ui(tab_gapji, "PMI")
        self._apply_sash_ratio("PMI")
        
        self.template_file_path.trace_add("write", lambda *a: self.load_template_specific_config("PMI"))

    def _update_pmi_ratio(self):
        """Saves current sash position and persists it."""
        try:
            total_w = self.pmi_paned.winfo_width()
            if total_w > 100:
                current_sash = self.pmi_paned.sash_coord(0)[0]
                self.pmi_pane_ratio = current_sash / total_w
                self.config['PMI_SASH_RATIO'] = self.pmi_pane_ratio
                self.save_settings() # Persist immediately
        except: pass

    def _on_pmi_paned_configure(self, event):
        """Maintains the proportional split when the window resizes."""
        try:
            total_w = self.pmi_paned.winfo_width()
            if total_w > 100:
                # During window resize, RE-PLACE sash based on SAVED ratio
                if event and event.widget == self.pmi_paned:
                    new_pos = int(total_w * self.pmi_pane_ratio)
                    self.pmi_paned.sash_place(0, new_pos, 0)
        except: pass

    def _update_pmi_filter_ui(self):
        """Redraws the element filter tags in a modern, elastic style."""
        for widget in self.filter_container.winfo_children():
            widget.destroy()
        
        def remove_filter(idx):
            self.element_filters.pop(idx)
            self._update_pmi_filter_ui()

        # [HYPER-REACTIVE] Use grid for even horizontal distribution (Elastic, Stacked)
        cols_per_row = 1
        for i, f_item in enumerate(self.element_filters):
            self.filter_container.columnconfigure(0, weight=1)
            
            # Create an "Elastic Tag" 
            tag = tk.Frame(self.filter_container, background="#eff6ff", highlightthickness=1, 
                            highlightbackground="#bfdbfe", padx=2, pady=1)
            tag.grid(row=i, column=0, sticky='ew', padx=2, pady=2)
            
            # Inner grid for tag components
            tag.columnconfigure(0, weight=1, minsize=0) # Key
            tag.columnconfigure(2, weight=1, minsize=0) # Min
            tag.columnconfigure(4, weight=1, minsize=0) # Max
            
            # Key
            e_key = tk.Entry(tag, textvariable=f_item['key'], width=1, font=("Arial", 8), 
                             relief='flat', background="#eff6ff", justify='center')
            e_key.grid(row=0, column=0, sticky='ew')
            tk.Label(tag, text=":", background="#eff6ff", font=("Arial", 8)).grid(row=0, column=1)
            
            # Min
            e_min = tk.Entry(tag, textvariable=f_item['min'], width=1, font=("Arial", 8), 
                             relief='flat', background="#eff6ff", justify='center')
            e_min.grid(row=0, column=2, sticky='ew')
            tk.Label(tag, text="~", background="#eff6ff", font=("Arial", 8)).grid(row=0, column=3)
            
            # Max
            e_max = tk.Entry(tag, textvariable=f_item['max'], width=1, font=("Arial", 8), 
                             relief='flat', background="#eff6ff", justify='center')
            e_max.grid(row=0, column=4, sticky='ew')
            
            # Delete button (Small)
            btn_del = tk.Button(tag, text="×", command=lambda idx=i: remove_filter(idx), 
                                relief='flat', background="#eff6ff", foreground="#3b82f6", font=("Arial", 9, "bold"),
                                activebackground="#dbeafe", cursor="hand2", padx=0, pady=0)
            btn_del.grid(row=0, column=5, padx=(2, 0))

    def _setup_rt_ui(self, parent):
        """[RE-INTEGRATED] RT 탭 통합 구성 (내부에 서브 미리보기 탭 배치)"""
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # --- Dual Pane Layout (PanedWindow) ---
        self.rt_paned = tk.PanedWindow(container, orient='horizontal', background="#d1d5db", 
                            sashwidth=6, sashpad=0, sashrelief='raised', borderwidth=0)
        self.rt_paned.pack(fill='both', expand=True)

        # [LEFT] Settings Sidebar
        left_container = tk.Frame(self.rt_paned, background="#f9fafb", highlightthickness=0, borderwidth=0)
        self.rt_paned.add(left_container, width=425)
        left_pane = self._create_scrollable_sidebar(left_container)

        # FIXED Header
        header_frame = tk.Frame(left_container, background="#f9fafb")
        header_frame.place(x=0, y=0, relwidth=1, height=40)
        tk.Label(header_frame, text="🔬 RT 성적서 통합 관리", font=("Malgun Gothic", 15, "bold"), 
                 background="#f9fafb", foreground="#1e3a8a").pack(side='left', padx=20, pady=5)
        tk.Frame(header_frame, height=1, background="#e5e7eb").pack(side='bottom', fill='x')
        tk.Frame(left_pane, height=40, background="#f9fafb").pack(fill='x')

        # [NEW] 모드 선택 (표준 vs 가스공사)
        mode_select_f = tk.Frame(left_pane, background="#f9fafb")
        mode_select_f.pack(fill='x', pady=(0, 10))
        self.rt_sub_mode = tk.StringVar(value="RT")
        
        style = ttk.Style()
        style.configure("Mode.TRadiobutton", font=("Malgun Gothic", 9, "bold"))
        
        ttk.Radiobutton(mode_select_f, text=" 🔬 표준 RT ", variable=self.rt_sub_mode, value="RT", 
                        style="Mode.TRadiobutton", command=self._on_rt_sub_mode_change).pack(side='left', padx=5)
        ttk.Radiobutton(mode_select_f, text=" 🇰🇷 가스공사 ", variable=self.rt_sub_mode, value="KOGAS", 
                        style="Mode.TRadiobutton", command=self._on_rt_sub_mode_change).pack(side='left', padx=5)

        # 1. File Selection (Adaptive)
        self.rt_file_frame = ttk.LabelFrame(left_pane, text=" 데이터 및 양식 (Data) ", padding=10)
        self.rt_file_frame.pack(fill='x', pady=(0, 10))
        
        # We need to show different path entries based on mode
        self.rt_path_container = tk.Frame(self.rt_file_frame)
        self.rt_path_container.pack(fill='x')
        
        def _add_file_row(p, lbl, var, r, is_dir=False, types=None):
            p.columnconfigure(1, weight=1)
            ttk.Label(p, text=lbl, font=("Malgun Gothic", 8)).grid(row=r, column=0, sticky='e', padx=2, pady=2)
            ttk.Entry(p, textvariable=var, font=("Arial", 9), width=1, exportselection=False).grid(row=r, column=1, padx=2, pady=2, sticky='ew')
            cmd = (lambda: self._browse_dir(var)) if is_dir else (lambda: self._browse_file(var, types))
            ttk.Button(p, text="...", width=3, command=cmd).grid(row=r, column=2, padx=2, pady=2)

        # Container for actual path rows (will be cleared/re-added on mode change)
        self._update_rt_path_ui()

        # 2. Tabs for Configuration
        rt_config_frame = ttk.LabelFrame(left_pane, text=" 세부 설정 (Config) ", padding=2)
        rt_config_frame.pack(fill='both', expand=True, pady=(0, 10))

        self.rt_tab_nb = ttk.Notebook(rt_config_frame)
        self.rt_tab_nb.pack(fill='both', expand=True)

        self.rt_tabs = {}
        for t_name in ["갑지", "을지", "로고", "행 설정", "컬럼 설정"]:
            f = ttk.Frame(self.rt_tab_nb, padding=5)
            self.rt_tab_nb.add(f, text=t_name)
            self.rt_tabs[t_name] = f
            if t_name not in ["행 설정", "컬럼 설정"]: f.columnconfigure(0, weight=1)

        self.rt_tab_nb.bind("<<NotebookTabChanged>>", self._update_gapji_preview_current)
        self._refresh_rt_config_tabs() # Initial UI build

        # 3. Actions
        action_f = tk.Frame(left_pane, background="#ffffff", highlightthickness=1, highlightbackground="#d1d5db", padx=10, pady=5)
        action_f.pack(fill='x', pady=(0, 10))
        
        # Row 1: Sequence Filter
        r1 = tk.Frame(action_f, background="#ffffff")
        r1.pack(fill='x', pady=2)
        ttk.Label(r1, text="📊 특정순번 필터:", background="#ffffff").pack(side='left')
        ttk.Entry(r1, textvariable=self.sequence_filter, width=15).pack(side='left', padx=5, fill='x', expand=True)
        
        # Row 2: Keyword Filter
        r2 = tk.Frame(action_f, background="#ffffff")
        r2.pack(fill='x', pady=2)
        ttk.Label(r2, text="🔍 추출 키워드:", background="#ffffff").pack(side='left')
        ttk.Entry(r2, textvariable=self.rt_extract_keyword, width=15).pack(side='left', padx=5, fill='x', expand=True)

        btn_r = tk.Frame(left_pane, background="#f9fafb")
        btn_r.pack(fill='x', pady=5)
        ttk.Button(btn_r, text=" ✨ 성적서 생성 ", style="Action.TButton", command=self.run_process).pack(fill='x', pady=(0, 5))
        ttk.Button(btn_r, text=" 📝 데이터 추출 ", command=self.extract_only).pack(fill='x', pady=(0, 5))
        ttk.Button(btn_r, text=" 🔄 의뢰서에 결과 반영 ", command=self.sync_results_to_request).pack(fill='x')

        # [RIGHT] Multi-Preview (Sub-tabs)
        right_f = tk.Frame(self.rt_paned, background="#f3f4f6")
        self.rt_paned.add(right_f, stretch="always")
        
        self.rt_preview_nb = ttk.Notebook(right_f)
        self.rt_preview_nb.pack(fill='both', expand=True)
        
        t_std = tk.Frame(self.rt_preview_nb, background="#ffffff")
        t_kogas = tk.Frame(self.rt_preview_nb, background="#ffffff")
        t_gapji = tk.Frame(self.rt_preview_nb, background="#ffffff")
        
        self.rt_preview_nb.add(t_std, text=" 🔬 표준 미리보기 ")
        self.rt_preview_nb.add(t_kogas, text=" 🇰🇷 가스공사 미리보기 ")
        self.rt_preview_nb.add(t_gapji, text=" 📄 갑지 미리보기 ")
        
        # Link Preview Tab to Mode Switch
        def _on_rt_preview_tab_changed(event):
            tab_text = self.rt_preview_nb.tab(self.rt_preview_nb.select(), "text")
            if "가스공사" in tab_text: self.rt_sub_mode.set("KOGAS")
            elif "표준" in tab_text: self.rt_sub_mode.set("RT")
            self._on_rt_sub_mode_change(from_tab=True)
        self.rt_preview_nb.bind("<<NotebookTabChanged>>", _on_rt_preview_tab_changed)

        self._create_rt_preview_ui(t_std, mode="RT")
        self._create_rt_preview_ui(t_kogas, mode="KOGAS")
        self._create_gapji_preview_ui(t_gapji, "RT") # Gapji uses current sub-mode context internally
        
        self._apply_sash_ratio("RT")
        self.rt_pane_ratio = self.config.get('RT_SASH_RATIO', 0.5)
        self.rt_paned.bind("<Configure>", lambda e: [self._on_rt_paned_configure(e), self.root.update_idletasks()])
        self.rt_paned.bind("<ButtonRelease-1>", lambda e: self.root.after(10, self._update_rt_ratio))
        self.root.after(500, lambda: self._on_rt_paned_configure(None))

    def _update_rt_path_ui(self):
        """모드에 따라 파일 선택 경로 입력 UI 갱신"""
        for w in self.rt_path_container.winfo_children(): w.destroy()
        mode = self.rt_sub_mode.get()
        v_target = self.rt_target_file_path if mode == "RT" else self.kogas_target_file_path
        v_template = self.rt_template_file_path if mode == "RT" else self.kogas_template_file_path

        def _add_file_row(p, lbl, var, r, is_dir=False, types=None):
            p.columnconfigure(1, weight=1)
            ttk.Label(p, text=lbl, font=("Malgun Gothic", 8)).grid(row=r, column=0, sticky='e', padx=2, pady=2)
            ttk.Entry(p, textvariable=var, font=("Arial", 9), width=1, exportselection=False).grid(row=r, column=1, padx=2, pady=2, sticky='ew')
            cmd = (lambda: self._browse_dir(var)) if is_dir else (lambda: self._browse_file(var, types))
            ttk.Button(p, text="...", width=3, command=cmd).grid(row=r, column=2, padx=2, pady=2)

        _add_file_row(self.rt_path_container, "기본 로고 폴더:", self.logo_folder_path, 0, is_dir=True)
        _add_file_row(self.rt_path_container, f"{mode} 데이터:", v_target, 1, types=[("Excel Source", "*.xls;*.xlsx;*.xlsm")])
        _add_file_row(self.rt_path_container, f"{mode} 양식:", v_template, 2, types=[("Excel Template", "*.xlsx;*.xlsm")])
        
        btn_f = tk.Frame(self.rt_path_container, background="#f0f0f0")
        btn_f.grid(row=3, column=1, sticky='ew', pady=2)
        ttk.Button(btn_f, text="💾 전용 설정 저장", command=lambda: self.save_template_specific_config(mode)).pack(side='left', padx=2)

    def _refresh_rt_config_tabs(self):
        """모드에 따라 행/컬럼 설정 탭 내용 갱신"""
        mode = self.rt_sub_mode.get()
        t_cover = self.rt_tabs["갑지"]
        t_data = self.rt_tabs["을지"]
        t_logo = self.rt_tabs["로고"]
        t_rows = self.rt_tabs["행 설정"]
        t_cols = self.rt_tabs["컬럼 설정"]
        
        for t in [t_rows, t_cols, t_logo, t_cover, t_data]:
            for w in t.winfo_children(): w.destroy()
        
        # Meta & Margins (Cover/Data)
        self._create_gapji_meta_ui(t_cover, use_pack=False)
        ctx_cover = f"{mode}_COVER"
        ctx_data = f"{mode}_DATA"
        self._create_margin_settings(t_cover, ctx_cover, use_pack=False)
        self._create_margin_settings(t_data, ctx_data, use_pack=False)
        
        # Logos
        tk.Label(t_logo, text="🚩 [갑지] 로고", font=("Malgun Gothic", 9, "bold"), background="#f9fafb").grid(row=0, column=0, sticky='w')
        r_next = self._create_setting_grid(t_logo, ctx_cover)
        tk.Label(t_logo, text="🚩 [을지] 로고", font=("Malgun Gothic", 9, "bold"), background="#f9fafb").grid(row=r_next+1, column=0, sticky='w', pady=(10, 0))
        self._create_setting_grid(t_logo, ctx_data)
        
        # Row Settings
        self._create_row_settings(t_rows, mode=mode)
        
        # Column Mapping (Ordered by current preview keys)
        all_possible = {
            "No": ("No:", f"{mode}_COL_NO", 1, f"{mode}_NAME_NO", "No", "No"),
            "Date": ("Date:", f"{mode}_COL_DATE", 2, f"{mode}_NAME_DATE", "Date", "Date"),
            "Dwg": ("Dwg No.:", f"{mode}_COL_DWG", 3, f"{mode}_NAME_DWG", "Drawing No.", "Dwg"),
            "Joint": ("Film No.:", f"{mode}_COL_JOINT", 4, f"{mode}_NAME_JOINT", "Film Ident. No.", "Joint"),
            "Loc": ("Location:", f"{mode}_COL_LOC", 5, f"{mode}_NAME_LOC", "Film Location", "Loc"),
            "T": ("T:", f"{mode}_COL_THK", 6, f"{mode}_NAME_THK", "T", "T"),
            "Mat": ("Mat:", f"{mode}_COL_MAT", 7, f"{mode}_NAME_MAT", "Mat", "Mat"),
            "Size": ("구경(Size):", f"{mode}_COL_SIZE", 0, f"{mode}_NAME_SIZE", "Size", "Size"),
            "Acc": ("Acc:", f"{mode}_COL_ACC", 8, f"{mode}_NAME_ACC", "Acc", "Acc"),
            "Rej": ("Rej:", f"{mode}_COL_REJ", 9, f"{mode}_NAME_REJ", "Rej", "Rej"),
            "Deg": ("Deg:", f"{mode}_COL_DEG", 10, f"{mode}_NAME_DEG", "Deg", "Deg"),
            "D1": ("① Crack:", f"{mode}_COL_D1", 13, f"{mode}_NAME_D1", "① Crack", "D1"),
            "D2": ("② IP:", f"{mode}_COL_D2", 14, f"{mode}_NAME_D2", "② IP", "D2"),
            "D3": ("③ LF:", f"{mode}_COL_D3", 15, f"{mode}_NAME_D3", "③ LF", "D3"),
            "D4": ("④ Slag:", f"{mode}_COL_D4", 16, f"{mode}_NAME_D4", "④ Slag", "D4"),
            "D5": ("⑤ Por:", f"{mode}_COL_D5", 17, f"{mode}_NAME_D5", "⑤ Por", "D5"),
            "⑥ U/C": ("⑥ U/C:", f"{mode}_COL_D6", 18, f"{mode}_NAME_D6", "⑥ U/C", "D6"),
            "D6": ("⑥ U/C:", f"{mode}_COL_D6", 18, f"{mode}_NAME_D6", "⑥ U/C", "D6"),
            "D7": ("⑦ RUC:", f"{mode}_COL_D7", 19, f"{mode}_NAME_D7", "⑦ RUC", "D7"),
            "D8": ("⑧ BT:", f"{mode}_COL_D8", 20, f"{mode}_NAME_D8", "⑧ BT", "D8"),
            "D9": ("⑨ TI:", f"{mode}_COL_D9", 21, f"{mode}_NAME_D9", "⑨ TI", "D9"),
            "D10": ("⑩ CP:", f"{mode}_COL_D10", 22, f"{mode}_NAME_D10", "⑩ CP", "D10"),
            "D11": ("⑪ RC:", f"{mode}_COL_D11", 23, f"{mode}_NAME_D11", "⑪ RC", "D11"),
            "D12": ("⑫ Mis:", f"{mode}_COL_D12", 24, f"{mode}_NAME_D12", "⑫ Mis", "D12"),
            "D13": ("⑬ EP:", f"{mode}_COL_D13", 25, f"{mode}_NAME_D13", "⑬ EP", "D13"),
            "D14": ("⑭ SD:", f"{mode}_COL_D14", 26, f"{mode}_NAME_D14", "⑭ SD", "D14"),
            "D15": ("⑮ Oth:", f"{mode}_COL_D15", 27, f"{mode}_NAME_D15", "⑮ Oth", "D15"),
            "Result": ("판정(Result):", f"{mode}_COL_RES", 28, f"{mode}_NAME_RES", "Result", "Result"),
            "Welder": ("용접사:", f"{mode}_COL_WELDER", 29, f"{mode}_NAME_WELDER", "Welder No", "Welder"),
            "Remarks": ("비고:", f"{mode}_COL_REM", 30, f"{mode}_NAME_REM", "Remarks", "Remarks")
        }
        
        keys_attr = "rt_column_keys" if mode == "RT" else "kogas_column_keys"
        current_keys = getattr(self, keys_attr)
        
        # 1. Start with items that are currently visible/ordered
        rt_items = []
        for k in current_keys:
            if k in all_possible:
                rt_items.append(all_possible[k])
        
        # 2. Add remaining items at the end (so they remain configurable even if hidden)
        for k, v in all_possible.items():
            if v not in rt_items:
                rt_items.append(v)

        if mode == "KOGAS":
            self._create_kogas_separated_column_mapping_ui(t_cols)
        else:
            self._create_column_mapping_ui(t_cols, mode, rt_items)

    def _on_rt_sub_mode_change(self, from_tab=False):
        """RT 내부 서브 모드(표준/가스공사) 변경 시 UI 및 상태 동기화"""
        mode = self.rt_sub_mode.get()
        self.current_mode = mode
        self.log(f"🔄 RT 서브 모드 전환: {mode}")
        
        # 1. 오른쪽 미리보기 탭 동기화 (라디오 버튼으로 바꾼 경우)
        if not from_tab:
            target_idx = 0 if mode == "RT" else 1
            if self.rt_preview_nb.index("current") != target_idx:
                self.rt_preview_nb.select(target_idx)
        
        # 2. 왼쪽 사이드바 UI 갱신 (파일 경로 및 세부 설정 탭)
        self._update_rt_path_ui()
        self._refresh_rt_config_tabs()
        
        # 3. 템플릿 로드 트리거 [FIX] Pass correct path and avoid forced tab switch
        t_path = self.rt_template_file_path.get() if mode == "RT" else self.kogas_template_file_path.get()
        self.load_template_specific_config(t_path, mode)
        
        # 4. 갑지 미리보기 갱신
        self._update_gapji_preview(mode)

    def _update_rt_ratio(self):
        try:
            total_w = self.rt_paned.winfo_width()
            if total_w > 100:
                current_sash = self.rt_paned.sash_coord(0)[0]
                self.rt_pane_ratio = current_sash / total_w
                self.config['RT_SASH_RATIO'] = self.rt_pane_ratio
                self.save_settings()
        except: pass

    def _on_rt_paned_configure(self, event):
        try:
            total_w = self.rt_paned.winfo_width()
            if total_w > 100:
                if event and event.widget == self.rt_paned:
                    new_pos = int(total_w * self.rt_pane_ratio)
                    self.rt_paned.sash_place(0, new_pos, 0)
        except: pass

    def _on_rt_kogas_mode_change(self):
        """가스공사 모드 변경 시 템플릿 경로 자동 설정 및 미리보기 컬럼 업데이트"""
        if self.rt_kogas_mode.get():
            # 가스공사 템플릿 경로 탐색
            data_dir = os.path.join(BASE_DIR, "Na-aba", "home", "data")
            if not os.path.exists(data_dir):
                data_dir = os.path.join(os.getcwd(), "Na-aba", "home", "data")
            
            kogas_template = os.path.join(data_dir, "가스공사 의뢰서.xlsx")
            if os.path.exists(kogas_template):
                self.rt_template_file_path.set(kogas_template)
                self.log("💡 가스공사 전용 템플릿이 자동으로 선택되었습니다.")
            else:
                self.log("⚠️ 가스공사 의뢰서.xlsx 파일을 찾을 수 없습니다. (data 폴더 확인 필요)")
        else:
            # 일반 KS 양식으로 복구 시도 (있는 경우)
            data_dir = os.path.join(BASE_DIR, "Na-aba", "home", "data")
            if not os.path.exists(data_dir):
                data_dir = os.path.join(os.getcwd(), "Na-aba", "home", "data")
                
            ks_template = os.path.join(data_dir, "RT KS양식.xlsx")
            if os.path.exists(ks_template):
                self.rt_template_file_path.set(ks_template)
                self.log("💡 일반 KS 양식 템플릿으로 복구되었습니다.")
        
        # [NEW] 미리보기 컬럼 구성 업데이트
        self._update_rt_preview_columns()

    def _update_rt_preview_columns(self, mode="RT"):
        """모드 여부에 따라 RT 미리보기 컬럼을 동적으로 변경 (사용자 설정 키 준수)"""
        tree = self.rt_preview_tree if mode == "RT" else self.kogas_preview_tree
        if not tree: return
        
        # [DYNAMIC] 고정된 리스트 대신 사용자가 설정한 키(rt_column_keys 등)를 사용
        keys_attr = "rt_column_keys" if mode == "RT" else "kogas_column_keys"
        current_cols = list(getattr(self, keys_attr))
            
        tree["columns"] = tuple(current_cols)
        
        # [NEW] 기본 표시 이름 맵핑
        default_names = {
            "selected": "V", "No": "No", "Date": "Date", "Dwg": "Drawing No.", 
            "Joint": "Film Ident. No.", "Loc": "Film Location", "T": "T", "Mat": "Mat",
            "Acc": "Acc", "Rej": "Rej", "Deg": "Deg", "Welder": "Welder No", "Remarks": "Remarks",
            "Dwg_Sub": "Dwg(Sub)", "Welder_Sub": "Welder(Sub)", "Mat_Sub": "Mat(Sub)",
            "D1": "① Crack", "D2": "② IP", "D3": "③ LF", "D4": "④ Slag", "D5": "⑤ Por",
            "D6": "⑥ U/C", "D7": "⑦ RUC", "D8": "⑧ BT", "D9": "⑨ TI", "D10": "⑩ CP",
            "D11": "⑪ RC", "D12": "⑫ Mis", "D13": "⑬ EP", "D14": "⑭ SD", "D15": "⑮ Oth"
        }
        
        saved_widths = self.config.get(f"{mode}_COL_WIDTHS", {})
        default_widths = {
            "selected": 40, "No": 50, "Date": 90, "Dwg": 300, "Joint": 120, "Loc": 100, 
            "Acc": 40, "Rej": 40, "Deg": 40, "Welder": 100, "Remarks": 120,
            "Dwg_Sub": 200, "Welder_Sub": 100, "Mat_Sub": 100
        }
        
        # [NEW] 내부 컬럼 ID와 설정 Key(rt_items 정의) 간의 맵핑
        key_map = {
            "T": "THK", "Remarks": "REM", "Result": "RES"
        }
        
        for col in current_cols:
            suffix = key_map.get(col, col.upper())
            name_key = f"{mode}_NAME_{suffix}"
            if col == "selected": name_key = None
            
            display_text = None
            if name_key:
                # 1. 현재 UI 입력값 확인 (최우선)
                if name_key in self.setting_vars:
                    display_text = self.setting_vars[name_key].get()
                
                # 2. 설정 파일(config) 확인
                if not display_text:
                    display_text = self.config.get(name_key)
                
                # 3. 보조 키 형식 확인 (소문자 등 대비)
                if not display_text:
                    alt_key = f"{mode}_NAME_{col}"
                    if alt_key in self.setting_vars:
                        display_text = self.setting_vars[alt_key].get()
                    else:
                        display_text = self.config.get(alt_key)

            fallback_name = default_names.get(col, col)
            if not display_text:
                display_text = fallback_name
                
            tree.heading(col, text=display_text, anchor='center', command=lambda _c=col: self.sort_by_column(_c, mode=mode))
            w = saved_widths.get(col, default_widths.get(col, 80))
            tree.column(col, width=w, anchor='center', stretch=False)
        
        # 데이터가 이미 있다면 리프레시
        data = self.rt_extracted_data if mode == "RT" else self.kogas_extracted_data
        if data:
            self.populate_preview(data, mode=mode, switch_tab=False)

    def _setup_pt_ui(self, parent):
        # [FORCE] Ensure parent (pt_mode_frame) allows expansion
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # --- Dual Pane Layout (PanedWindow) ---
        self.pt_paned = tk.PanedWindow(container, orient='horizontal', background="#d1d5db", 
                            sashwidth=6, sashpad=0, sashrelief='raised', borderwidth=0)
        self.pt_paned.pack(fill='both', expand=True)

        # [LEFT] Settings Sidebar
        left_container = tk.Frame(self.pt_paned, background="#f9fafb", highlightthickness=0, borderwidth=0)
        self.pt_paned.add(left_container, width=425)
        
        # Scrollable area (Full size)
        left_pane = self._create_scrollable_sidebar(left_container)

        # FIXED FLOATING Header
        header_frame = tk.Frame(left_container, background="#f9fafb", highlightthickness=0, borderwidth=0)
        header_frame.place(x=0, y=0, relwidth=1, height=40)
        
        inner_header = tk.Frame(header_frame, background="#f9fafb", padx=20)
        inner_header.pack(fill='both', expand=True, pady=(5, 0))
        
        tk.Label(inner_header, text="🔬 PT 성적서 관리", font=("Malgun Gothic", 15, "bold"), 
                 background="#f9fafb", foreground="#1e3a8a").pack(side='left')
        
        tk.Frame(header_frame, height=1, background="#e5e7eb").pack(side='bottom', fill='x')
        header_frame.lift()

        # Spacer in scrollable area
        tk.Frame(left_pane, height=40, background="#f9fafb").pack(fill='x')

        # 1. File Selection Group
        file_frame = ttk.LabelFrame(left_pane, text=" 데이터 및 양식 (Data) ", padding=10)
        file_frame.pack(fill='x', pady=(0, 10))

        def _add_file_row(parent_f, label, var, row, is_dir=False, types=None):
            parent_f.columnconfigure(1, weight=1)
            ttk.Label(parent_f, text=label, font=("Malgun Gothic", 8)).grid(row=row, column=0, sticky='e', padx=2, pady=2)
            ttk.Entry(parent_f, textvariable=var, font=("Arial", 9), width=1, exportselection=False).grid(row=row, column=1, padx=2, pady=2, sticky='ew')
            cmd = (lambda: self._browse_dir(var)) if is_dir else (lambda: self._browse_file(var, types))
            ttk.Button(parent_f, text="...", width=3, command=cmd).grid(row=row, column=2, padx=2, pady=2)

        _add_file_row(file_frame, "기본 로고 폴더:", self.logo_folder_path, 0, is_dir=True)
        _add_file_row(file_frame, "PT 데이터:", self.pt_target_file_path, 1, types=[("Excel Source", "*.xls;*.xlsx;*.xlsm")])
        _add_file_row(file_frame, "PT 양식:", self.pt_template_file_path, 2, types=[("Excel Template", "*.xlsx;*.xlsm")])
        
        # [NEW] Template Config Save Button
        btn_frame = tk.Frame(file_frame, background="#f0f0f0")
        btn_frame.grid(row=3, column=1, sticky='ew', pady=2)
        ttk.Button(btn_frame, text="💾 현재 설정을 이 양식 전용으로 저장", command=lambda: self.save_template_specific_config("PT")).pack(side='left', padx=2)

        # 2. Configuration Tabs
        pt_config_frame = ttk.LabelFrame(left_pane, text=" 리포트 세부 설정 ", padding=2)
        pt_config_frame.pack(fill='both', expand=True, pady=(0, 10))

        self.pt_tab_notebook = ttk.Notebook(pt_config_frame)
        self.pt_tab_notebook.pack(fill='both', expand=True)

        pt_tab_cover = ttk.Frame(self.pt_tab_notebook, padding=5)
        pt_tab_data = ttk.Frame(self.pt_tab_notebook, padding=5)
        pt_tab_rows = ttk.Frame(self.pt_tab_notebook, padding=5)
        pt_tab_cols = ttk.Frame(self.pt_tab_notebook, padding=5)
        
        # [CRITICAL] Allow children to expand
        for t in [pt_tab_cover, pt_tab_data, pt_tab_rows, pt_tab_cols]: t.columnconfigure(0, weight=1)

        self.pt_tab_notebook.add(pt_tab_cover, text="갑지")
        self.pt_tab_notebook.add(pt_tab_data, text="을지")
        self.pt_tab_notebook.add(pt_tab_rows, text="행 설정")
        self.pt_tab_notebook.add(pt_tab_cols, text="컬럼 설정")

        # [ALIGNED] Mode-specific context for logo grid
        self._create_gapji_meta_ui(pt_tab_cover, use_pack=False)
        self.pt_tab_notebook.bind("<<NotebookTabChanged>>", self._update_gapji_preview_current)
        next_row_pt_cover = self._create_setting_grid(pt_tab_cover, "PT_COVER")
        next_row_pt_data = self._create_setting_grid(pt_tab_data, "PT_DATA")
        self._create_margin_settings(pt_tab_cover, "PT_COVER", use_pack=False)
        self._create_margin_settings(pt_tab_data, "PT_DATA", use_pack=False)
        self._create_row_settings(pt_tab_rows, mode="PT")
        
        pt_items = [
            ("순번(No):", "PT_COL_NO", 1, "PT_NAME_NO", "No", "No"),
            ("ISO/Dwg:", "PT_COL_DWG", 2, "PT_NAME_DWG", "Dwg", "Dwg"),
            ("Joint No:", "PT_COL_JOINT", 5, "PT_NAME_JOINT", "Joint", "Joint"),
            ("NPS:", "PT_COL_NPS", 6, "PT_NAME_NPS", "NPS", "NPS"),
            ("두께(Th'k):", "PT_COL_THK", 7, "PT_NAME_THK", "Thk.", "Thk."),
            ("재질(Material):", "PT_COL_MAT", 8, "PT_NAME_MAT", "Material", "Material"),
            ("용접사(Welder):", "PT_COL_WELDER", 9, "PT_NAME_WELDER", "Welder", "Welder"),
            ("검사타입(Type):", "PT_COL_TYPE", 10, "PT_NAME_TYPE", "WType", "WType"),
            ("판정 결과:", "PT_COL_RES", 11, "PT_NAME_RES", "Result", "Result")
        ]
        self._create_column_mapping_ui(pt_tab_cols, "PT", pt_items)
        

        # 3. Action Section
        action_frame = tk.Frame(left_pane, background="#ffffff", highlightthickness=1, highlightbackground="#d1d5db", padx=10, pady=5)
        action_frame.pack(fill='x', pady=(0, 10))

        filter_row = tk.Frame(action_frame, background="#ffffff")
        filter_row.pack(fill='x', pady=2)
        ttk.Label(filter_row, text="📊 특정순번:", background="#ffffff", font=("Malgun Gothic", 8)).pack(side='left')
        ttk.Entry(filter_row, textvariable=self.sequence_filter, width=15, font=("Arial", 9), exportselection=False).pack(side='left', padx=5, fill='x', expand=True)

        btn_row = tk.Frame(left_pane, background="#f9fafb")
        btn_row.pack(fill='x', pady=5)
        ttk.Button(btn_row, text=" ✨ 성적서 생성 ", style="Action.TButton", command=self.run_process).pack(fill='x', pady=(0, 5))
        ttk.Button(btn_row, text=" 📝 데이터 추출 ", command=self.extract_only).pack(fill='x')

        # [RIGHT] Multi-Preview Pane (Data & Gapji)
        right_container = tk.Frame(self.pt_paned, background="#f3f4f6")
        self.pt_paned.add(right_container, stretch="always")
        
        if not hasattr(self, 'preview_notebooks'): self.preview_notebooks = {}
        nb = ttk.Notebook(right_container)
        nb.pack(fill='both', expand=True)
        self.preview_notebooks["PT"] = nb
        
        tab_data = tk.Frame(nb, background="#ffffff")
        tab_gapji = tk.Frame(nb, background="#ffffff")
        nb.add(tab_data, text=" 🔬 데이터 미리보기 ")
        nb.add(tab_gapji, text=" 📄 갑지 미리보기 ")
        
        self._create_pt_preview_ui(tab_data)
        self._create_gapji_preview_ui(tab_gapji, "PT")
        self._apply_sash_ratio("PT")
        
        self.pt_template_file_path.trace_add("write", lambda *a: self.load_template_specific_config("PT"))

        # Adaptive Resizing Bindings
        self.pt_pane_ratio = self.config.get('PT_SASH_RATIO', 0.5)
        self.pt_paned.bind("<Configure>", lambda e: [self._on_pt_paned_configure(e), self.root.update_idletasks()])
        self.pt_paned.bind("<ButtonRelease-1>", lambda e: self.root.after(10, self._update_pt_ratio))
        self.root.after(500, lambda: self._on_pt_paned_configure(None))

    def _update_pt_ratio(self):
        try:
            total_w = self.pt_paned.winfo_width()
            if total_w > 100:
                current_sash = self.pt_paned.sash_coord(0)[0]
                self.pt_pane_ratio = current_sash / total_w
                self.config['PT_SASH_RATIO'] = self.pt_pane_ratio
                self.save_settings()
        except: pass

    def _on_pt_paned_configure(self, event):
        try:
            total_w = self.pt_paned.winfo_width()
            if total_w > 100:
                if event and event.widget == self.pt_paned:
                    new_pos = int(total_w * self.pt_pane_ratio)
                    self.pt_paned.sash_place(0, new_pos, 0)
        except: pass

    def _create_pt_preview_ui(self, parent):
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # [NEW] File Info Header
        header_info = tk.Frame(container, background="#ffffff", highlightthickness=1, highlightbackground="#e5e7eb")
        header_info.pack(fill='x', pady=(0, 5))
        tk.Label(header_info, textvariable=self.file_info_vars['PT'], background="#ffffff", 
                 foreground="#4b5563", font=("Malgun Gothic", 8, "bold"), padx=10, pady=2).pack(side='left')

        self.pt_display_cols = ["V", "No", "Date", "ISO/Dwg", "Joint No.", "Material", "Test Item", "Result", "Welder No", "Remarks"]
        saved_widths = self.config.get("PT_COL_WIDTHS", {})
        default_widths = {"V": 40, "No": 50, "Date": 90, "ISO/Dwg": 300, "Joint No.": 120, "Material": 100, "Test Item": 100, "Result": 80, "Welder No": 100}

        tree_frame = tk.Frame(container, background="#f9fafb")
        self.pt_preview_tree = ttk.Treeview(tree_frame, columns=self.pt_display_cols, show='headings', height=10, selectmode='extended')
        for col in self.pt_preview_tree["columns"]:
            name_key = f"PT_NAME_{col.split('(')[0].replace(' ', '').replace('.', '').upper()}"
            if col == "Date": name_key = "PT_NAME_DATE"
            elif col == "ISO/Dwg": name_key = "PT_NAME_DWG"
            elif col == "Joint No.": name_key = "PT_NAME_JOINT"
            elif col == "Material": name_key = "PT_NAME_MAT"
            elif col == "Thk": name_key = "PT_NAME_THK"
            elif col == "Test Item": name_key = "PT_NAME_ITEM"
            elif col == "Result": name_key = "PT_NAME_RES"
            elif col == "Welder No": name_key = "PT_NAME_WELDER"
            else: name_key = None
            
            display_text = self.config.get(name_key, col) if name_key else col
            self.pt_preview_tree.heading(col, text=display_text, anchor='center', command=lambda _c=col: self.sort_by_column(_c, mode="PT"))
            w = saved_widths.get(col, default_widths.get(col, 80))
            self.pt_preview_tree.column(col, width=w, anchor='center', stretch=False)
        
        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.pt_preview_tree.yview)
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.pt_preview_tree.xview)
        self.pt_preview_tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.pt_preview_tree.grid(row=0, column=0, sticky='nsew')
        scroll_y.grid(row=0, column=1, sticky='ns')
        scroll_x.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        self._setup_preview_sidebar(self.pt_preview_tree, container, mode="PT")
        tree_frame.pack(side="left", fill="both", expand=True)

    def _setup_paut_ui(self, parent):
        # [FORCE] Ensure parent (paut_mode_frame) allows expansion
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # --- Dual Pane Layout (PanedWindow) ---
        self.paut_paned = tk.PanedWindow(container, orient='horizontal', background="#d1d5db", 
                            sashwidth=6, sashpad=0, sashrelief='raised', borderwidth=0)
        self.paut_paned.pack(fill='both', expand=True)

        # [LEFT] Settings & Manual Eval Pane
        left_container = tk.Frame(self.paut_paned, background="#f9fafb", highlightthickness=0, borderwidth=0)
        self.paut_paned.add(left_container, width=425)
        
        # Scrollable area (Full size)
        left_pane = self._create_scrollable_sidebar(left_container)

        # FIXED FLOATING Header
        header_frame = tk.Frame(left_container, background="#f9fafb", highlightthickness=0, borderwidth=0)
        header_frame.place(x=0, y=0, relwidth=1, height=40)
        
        inner_header = tk.Frame(header_frame, background="#f9fafb", padx=20)
        inner_header.pack(fill='both', expand=True, pady=(5, 0))
        
        tk.Label(inner_header, text="🔬 PAUT (ASME B31.1) 관리", font=("Malgun Gothic", 15, "bold"), 
                 background="#f9fafb", foreground="#1e3a8a").pack(side='left')
        
        tk.Frame(header_frame, height=1, background="#e5e7eb").pack(side='bottom', fill='x')
        header_frame.lift()

        # Spacer in scrollable area
        tk.Frame(left_pane, height=40, background="#f9fafb").pack(fill='x')

        # 1. File Selection Group
        file_frame = ttk.LabelFrame(left_pane, text=" 데이터 및 양식 (Data) ", padding=10)
        file_frame.pack(fill='x', pady=(0, 10))

        def _add_file_row(parent_f, label, var, row, is_dir=False, types=None):
            parent_f.columnconfigure(1, weight=1)
            ttk.Label(parent_f, text=label, font=("Malgun Gothic", 8)).grid(row=row, column=0, sticky='e', padx=2, pady=2)
            ttk.Entry(parent_f, textvariable=var, font=("Arial", 9), width=1, exportselection=False).grid(row=row, column=1, padx=2, pady=2, sticky='ew')
            cmd = (lambda: self._browse_dir(var)) if is_dir else (lambda: self._browse_file(var, types))
            ttk.Button(parent_f, text="...", width=3, command=cmd).grid(row=row, column=2, padx=2, pady=2)

        _add_file_row(file_frame, "기본 로고 폴더:", self.logo_folder_path, 0, is_dir=True)
        _add_file_row(file_frame, "PAUT 데이터:", self.paut_target_file_path, 1, types=[("Excel Source", "*.xls;*.xlsx;*.xlsm")])
        _add_file_row(file_frame, "PAUT 양식:", self.paut_template_file_path, 2, types=[("Excel Template", "*.xlsx;*.xlsm")])

        # --- 2. Configuration Notebook (갑, 을, 행 설정, 개별 판정) ---
        config_frame = ttk.LabelFrame(left_pane, text=" 설정 및 판정 (Config & Eval) ", padding=2)
        config_frame.pack(fill='both', expand=True, pady=(0, 10))

        self.paut_tab_notebook = ttk.Notebook(config_frame)
        self.paut_tab_notebook.pack(fill='both', expand=True)

        tab_cover = ttk.Frame(self.paut_tab_notebook, padding=5)
        tab_data = ttk.Frame(self.paut_tab_notebook, padding=5)
        tab_rows = ttk.Frame(self.paut_tab_notebook, padding=5)
        tab_cols = ttk.Frame(self.paut_tab_notebook, padding=5)
        tab_eval = ttk.Frame(self.paut_tab_notebook, padding=5)

        # Allow expansion
        for t in [tab_cover, tab_data, tab_rows, tab_cols, tab_eval]: t.columnconfigure(0, weight=1)

        self.paut_tab_notebook.add(tab_eval, text="개별 판정") # Default/Main tab
        self.paut_tab_notebook.add(tab_cover, text="갑지")
        self.paut_tab_notebook.add(tab_data, text="을지")
        self.paut_tab_notebook.add(tab_rows, text="행 설정")
        self.paut_tab_notebook.add(tab_cols, text="컬럼 설정")

        # [NEW] Populate Settings
        self._create_gapji_meta_ui(tab_cover, use_pack=False)
        self.paut_tab_notebook.bind("<<NotebookTabChanged>>", self._update_gapji_preview_current)
        next_row_cover = self._create_setting_grid(tab_cover, "PAUT_COVER")
        next_row_data = self._create_setting_grid(tab_data, "PAUT_DATA")
        self._create_margin_settings(tab_cover, "PAUT_COVER", use_pack=False)
        self._create_margin_settings(tab_data, "PAUT_DATA", use_pack=False)
        self._create_row_settings(tab_rows, mode="PAUT")
        
        paut_items = [
            ("순번(No):", "PAUT_COL_NO", 1, "PAUT_COL_NO_NAME", "No", "No"),
            ("Line No.:", "PAUT_COL_LINE", 2, "PAUT_NAME_LINE", "Line No.", "Line No."),
            ("Joint No.:", "PAUT_COL_JOINT", 3, "PAUT_NAME_JOINT", "Joint No.", "Joint No."),
            ("두께(Th'k):", "PAUT_COL_THK", 4, "PAUT_NAME_THK", "Th'k(mm)", "Th'k(mm)"),
            ("결함높이(H):", "PAUT_COL_H", 5, "PAUT_NAME_H", "Height(mm)", "Height(mm)"),
            ("결함길이(L):", "PAUT_COL_L", 6, "PAUT_NAME_L", "Length(mm)", "Length(mm)"),
            ("결함깊이(D):", "PAUT_COL_UP", 7, "PAUT_NAME_UP", "Upper", "Upper"),
            ("하부깊이(Low):", "PAUT_COL_LOW", 0, "PAUT_NAME_LOW", "Lower", "Lower"),
            ("시작위치:", "PAUT_COL_START", 0, "PAUT_NAME_START", "Start", "Start"),
            ("종료위치:", "PAUT_COL_END", 0, "PAUT_NAME_END", "End", "End"),
            ("a/l 비:", "PAUT_COL_AL", 0, "PAUT_NAME_AL", "a/l", "a/l"),
            ("a/t 비:", "PAUT_COL_AT", 0, "PAUT_NAME_AT", "a/t", "a/t"),
            ("결함종류:", "PAUT_COL_NAT", 8, "PAUT_NAME_NAT", "Type of Flaw", "Type of Flaw"),
            ("판정(Eval):", "PAUT_COL_EVAL", 9, "PAUT_NAME_EVAL", "Evaluation", "Evaluation"),
            ("비고(Rem):", "PAUT_COL_REM", 10, "PAUT_NAME_REM", "Remarks", "Remarks")
        ]
        self._create_column_mapping_ui(tab_cols, "PAUT", paut_items)

        # --- 3. Manual Evaluation Content (Inside tab_eval) ---
        manual_container = tk.Frame(tab_eval, background="#f9fafb")
        manual_container.pack(fill='both', expand=True)
        
        # [NEW] Evaluation Mode Selector
        mode_frame = tk.Frame(manual_container, background="#f9fafb")
        mode_frame.pack(fill='x', pady=(0, 5))
        ttk.Label(mode_frame, text="✅ 판정 기준:", font=("Malgun Gothic", 8, "bold")).pack(side='left', padx=2)
        tk.Radiobutton(mode_frame, text="일반(B31.1)", variable=self.paut_eval_mode, value="Standard", background="#f9fafb").pack(side='left', padx=5)
        tk.Radiobutton(mode_frame, text="ECA(파과역학)", variable=self.paut_eval_mode, value="ECA", background="#f9fafb").pack(side='left', padx=5)

        # Sizing Helper Row
        calc_frame = tk.Frame(manual_container, background="#f1f5f9", padx=10, pady=8, highlightbackground="#cbd5e1", highlightthickness=1)
        calc_frame.pack(fill='x', pady=(0, 10))
        
        # Row 0: H-Calc Settings
        h_set_row = tk.Frame(calc_frame, background="#f1f5f9")
        h_set_row.pack(fill='x')
        ttk.Label(h_set_row, text="📏 H-Setting:", font=("Malgun Gothic", 8, "bold"), background="#f1f5f9").pack(side='left', padx=2)
        tk.Entry(h_set_row, textvariable=self.paut_manual_vars['peak'], width=5).pack(side='left', padx=2)
        ttk.Label(h_set_row, text="% /", background="#f1f5f9").pack(side='left')
        tk.Entry(h_set_row, textvariable=self.paut_manual_vars['db'], width=4).pack(side='left', padx=2)
        ttk.Label(h_set_row, text="dB ->", background="#f1f5f9").pack(side='left', padx=2)
        ttk.Label(h_set_row, textvariable=self.paut_manual_vars['target_fsh'], background="#f1f5f9", foreground="#ef4444", font=("Arial", 8, "bold")).pack(side='left', padx=2)
        
        # Row 1: H-Calc Data
        h_data_row = tk.Frame(calc_frame, background="#f1f5f9")
        h_data_row.pack(fill='x', pady=2)
        ttk.Label(h_data_row, text="📏 H-Data:", font=("Malgun Gothic", 8, "bold"), background="#f1f5f9").pack(side='left', padx=2)
        tk.Entry(h_data_row, textvariable=self.paut_manual_vars['z1'], width=6).pack(side='left', padx=2)
        ttk.Label(h_data_row, text="~", background="#f1f5f9").pack(side='left')
        tk.Entry(h_data_row, textvariable=self.paut_manual_vars['z2'], width=6).pack(side='left', padx=2)
        ttk.Button(h_data_row, text="입력", width=6, command=self._calculate_paut_h).pack(side='right', padx=5)

        # Row 2: L-Calc Data
        l_data_row = tk.Frame(calc_frame, background="#f1f5f9")
        l_data_row.pack(fill='x')
        ttk.Label(l_data_row, text="📏 L-Data:", font=("Malgun Gothic", 8, "bold"), background="#f1f5f9").pack(side='left', padx=2)
        tk.Entry(l_data_row, textvariable=self.paut_manual_vars['L1'], width=6).pack(side='left', padx=2)
        ttk.Label(l_data_row, text="~", background="#f1f5f9").pack(side='left')
        tk.Entry(l_data_row, textvariable=self.paut_manual_vars['L2'], width=6).pack(side='left', padx=2)
        ttk.Button(l_data_row, text="입력", width=6, command=self._calculate_paut_l).pack(side='right', padx=5)

        # Input Grid
        input_grid = tk.Frame(manual_container, background="#f9fafb")
        input_grid.pack(fill='x', pady=5)
        
        m_inputs = [("두께(T):", "t"), ("높이(H):", "h"), ("길이(L):", "l"), ("깊이(d):", "d"), 
                    ("시작(S):", "L1"), ("종료(E):", "L2"), ("상단(Up):", "D_Upper"), ("하단(Low):", "D_Lower")]
        for i, (lbl, key) in enumerate(m_inputs):
            grid_row = i // 2
            grid_col = (i % 2) * 2
            ttk.Label(input_grid, text=lbl, font=("Malgun Gothic", 8)).grid(row=grid_row, column=grid_col, sticky='e', padx=2, pady=1)
            ent = tk.Entry(input_grid, textvariable=self.paut_manual_vars[key], width=10)
            ent.grid(row=grid_row, column=grid_col+1, padx=2, pady=1, sticky='w')
            ent.bind("<Return>", lambda e: self.root.focus_set())

        # Nature & Evaluation Actions
        nature_row = tk.Frame(manual_container, background="#f9fafb")
        nature_row.pack(fill='x', pady=5)
        ttk.Label(nature_row, text="종류:").pack(side='left', padx=2)
        ttk.Combobox(nature_row, textvariable=self.paut_manual_vars['nature'], values=["Crack", "LOF", "IP", "Slag", "Porosity", "Others"], width=9).pack(side='left', padx=2)
        ttk.Label(nature_row, text="위치:").pack(side='left', padx=2)
        ttk.Label(nature_row, textvariable=self.paut_manual_vars['loc'], foreground="#3b82f6", font=("Arial", 9, "bold")).pack(side='left', padx=2)
        
        btn_row = tk.Frame(manual_container, background="#f9fafb")
        btn_row.pack(fill='x', pady=2)
        ttk.Button(btn_row, text="판정", width=8, command=self._run_manual_paut_eval).pack(side='left', expand=True, fill='x', padx=2)
        ttk.Button(btn_row, text="저장", width=8, command=self._save_paut_manual_to_tree).pack(side='left', expand=True, fill='x', padx=2)

        self.paut_res_label = tk.Label(manual_container, text="결과 대기", font=("Malgun Gothic", 10, "bold"), background="#f3f4f6")
        self.paut_res_label.pack(fill='x', pady=(5, 0))

        # Bottom Actions
        paut_btn_row = tk.Frame(left_pane, background="#f9fafb")
        paut_btn_row.pack(fill='x', pady=(5, 2))
        ttk.Button(paut_btn_row, text=" 📝 추출 ", command=self._extract_paut_data).pack(side='left', fill='x', expand=True, padx=(0, 5))
        ttk.Button(paut_btn_row, text=" ✨ 일괄 판정 ", command=self._run_batch_paut_eval).pack(side='left', fill='x', expand=True, padx=(0, 5))
        ttk.Button(paut_btn_row, text=" 📄 성적서 생성 ", command=self._generate_paut_report).pack(side='left', fill='x', expand=True)

        paut_session_row = tk.Frame(left_pane, background="#f9fafb")
        paut_session_row.pack(fill='x', pady=(0, 10))
        ttk.Button(paut_session_row, text=" 💾 세션 저장 ", command=self._export_paut_session).pack(side='left', fill='x', expand=True, padx=(0, 5))
        ttk.Button(paut_session_row, text=" 📂 세션 불러오기 ", command=self._import_paut_session).pack(side='left', fill='x', expand=True)

        # [RIGHT] Multi-Preview Pane (Data & Gapji)
        right_container = tk.Frame(self.paut_paned, background="#f3f4f6")
        self.paut_paned.add(right_container, stretch="always")
        
        if not hasattr(self, 'preview_notebooks'): self.preview_notebooks = {}
        nb = ttk.Notebook(right_container)
        nb.pack(fill='both', expand=True)
        self.preview_notebooks["PAUT"] = nb
        
        tab_data = tk.Frame(nb, background="#ffffff")
        tab_gapji = tk.Frame(nb, background="#ffffff")
        nb.add(tab_data, text=" 🔬 데이터 미리보기 ")
        nb.add(tab_gapji, text=" 📄 갑지 미리보기 ")
        
        self._create_paut_preview_ui(tab_data)
        self._create_gapji_preview_ui(tab_gapji, "PAUT")
        self._apply_sash_ratio("PAUT")
        
    def _create_paut_preview_ui(self, parent):
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # [NEW] File Info Header
        header_info = tk.Frame(container, background="#ffffff", highlightthickness=1, highlightbackground="#e5e7eb")
        header_info.pack(fill='x', pady=(0, 5))
        tk.Label(header_info, textvariable=self.file_info_vars['PAUT'], background="#ffffff", 
                 foreground="#4b5563", font=("Malgun Gothic", 8, "bold"), padx=10, pady=2).pack(side='left')

        tree_frame = tk.Frame(container, background="#f9fafb")
        tree_frame.pack(fill='both', expand=True)
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        cols = ("V", "No", "Line No.", "Joint No.", "Th'k(mm)", "Start", "End", "Length(mm)", "Upper", "Lower", "Height(mm)", "Type of Flaw", "a/l", "a/t", "Evaluation", "Remarks")
        self.paut_preview_tree = ttk.Treeview(tree_frame, columns=cols, show='headings', height=10, selectmode='extended')
        
        saved_widths = self.config.get("PAUT_COL_WIDTHS", {})
        default_widths = {"V": 40, "No": 50, "Line No.": 250, "Joint No.": 120, "Th'k(mm)": 60, "Start": 60, "End": 60, "Length(mm)": 80, "Upper": 60, "Lower": 60, "Height(mm)": 80, "Type of Flaw": 100, "a/l": 60, "a/t": 60, "Evaluation": 80, "Remarks": 150}
        
        for col in self.paut_preview_tree["columns"]:
            # Use dynamic names from config for headings
            name_key = f"PAUT_NAME_{col.split('(')[0].replace(' ', '').replace('.', '').replace('/', '').upper()}"
            default_name = col
            if col == "No": name_key = "PAUT_COL_NO_NAME"
            elif col == "Line No.": name_key = "PAUT_NAME_LINE"
            elif col == "Joint No.": name_key = "PAUT_NAME_JOINT"
            elif col == "Th'k(mm)": name_key = "PAUT_NAME_THK"
            elif col == "Height(mm)": name_key = "PAUT_NAME_H"
            elif col == "Length(mm)": name_key = "PAUT_NAME_L"
            elif col == "Upper": name_key = "PAUT_NAME_UP"
            elif col == "Lower": name_key = "PAUT_NAME_LOW"
            elif col == "Start": name_key = "PAUT_NAME_START"
            elif col == "End": name_key = "PAUT_NAME_END"
            elif col == "a/l": name_key = "PAUT_NAME_AL"
            elif col == "a/t": name_key = "PAUT_NAME_AT"
            elif col == "Type of Flaw": name_key = "PAUT_NAME_NAT"
            elif col == "Evaluation": name_key = "PAUT_NAME_EVAL"
            elif col == "Remarks": name_key = "PAUT_NAME_REM"
            else: name_key = None
            
            display_text = self.config.get(name_key, default_name) if name_key else default_name
            self.paut_preview_tree.heading(col, text=display_text, anchor='center', command=lambda _c=col: self.sort_by_column(_c, mode="PAUT"))
            
            w = saved_widths.get(col, default_widths.get(col, 80))
            self.paut_preview_tree.column(col, width=w, anchor='center', stretch=False)
        
        self.paut_preview_tree.grid(row=0, column=0, sticky='nsew')
        
        paut_vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.paut_preview_tree.yview)
        paut_hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.paut_preview_tree.xview)
        self.paut_preview_tree.configure(yscrollcommand=paut_vsb.set, xscrollcommand=paut_hsb.set)
        
        paut_vsb.grid(row=0, column=1, sticky='ns')
        paut_hsb.grid(row=1, column=0, sticky='ew')

        self._setup_preview_sidebar(self.paut_preview_tree, container, mode="PAUT")


        # Adaptive Resizing Bindings
        self.paut_pane_ratio = self.config.get('PAUT_SASH_RATIO', 0.5)
        self.paut_paned.bind("<Configure>", lambda e: [self._on_paut_paned_configure(e), self.root.update_idletasks()])
        self.paut_paned.bind("<ButtonRelease-1>", lambda e: self.root.after(10, self._update_paut_ratio))
        self.root.after(500, lambda: self._on_paut_paned_configure(None))

    def _update_paut_ratio(self):
        try:
            total_w = self.paut_paned.winfo_width()
            if total_w > 100:
                current_sash = self.paut_paned.sash_coord(0)[0]
                self.paut_pane_ratio = current_sash / total_w
                self.config['PAUT_SASH_RATIO'] = self.paut_pane_ratio
                self.save_settings()
        except: pass

    def _on_paut_paned_configure(self, event):
        try:
            total_w = self.paut_paned.winfo_width()
            if total_w > 100:
                if event and event.widget == self.paut_paned:
                    new_pos = int(total_w * self.paut_pane_ratio)
                    self.paut_paned.sash_place(0, new_pos, 0)
        except: pass

    def _update_paut_auto_loc(self):
        try:
            t = float(self.paut_manual_vars['t'].get() or 0)
            d = float(self.paut_manual_vars['d'].get() or 0)
            h = float(self.paut_manual_vars['h'].get() or 0)
            if t > 0:
                is_surface = (d <= 0.1 * t) or (d + h >= 0.9 * t)
                loc = "Surface" if is_surface else "Subsurface"
                self.paut_manual_vars['loc'].set(loc)
        except: pass

    def _update_paut_target_fsh(self):
        """[NEW] Peak 나 dB 변경 시 목표 % FSH 실시간 업데이트"""
        try:
            peak = float(self.paut_manual_vars['peak'].get() or 0)
            db = float(self.paut_manual_vars['db'].get() or 0)
            if peak > 0:
                drop_factor = pow(10, -db / 20)
                target = peak * drop_factor
                self.paut_manual_vars['target_fsh'].set(f"{target:.1f}%")
            else:
                self.paut_manual_vars['target_fsh'].set("0.0%")
        except:
            self.paut_manual_vars['target_fsh'].set("-")

    def _calculate_paut_h(self):
        """[NEW] dB 강하법을 이용한 h 값 계산 및 자동 입력"""
        try:
            peak = float(self.paut_manual_vars['peak'].get() or 80)
            db = float(self.paut_manual_vars['db'].get() or 6)
            z1 = float(self.paut_manual_vars['z1'].get() or 0)
            z2 = float(self.paut_manual_vars['z2'].get() or 0)
            
            # Calculate Target FSH: Peak * 10^(-dB/20)
            drop_factor = pow(10, -db / 20)
            target_fsh = peak * drop_factor
            self.paut_manual_vars['target_fsh'].set(f"{target_fsh:.1f}%")

            h_calc = abs(z1 - z2)
            self.paut_manual_vars['h'].set(f"{h_calc:.2f}")
            
            # [NEW] D: 중간값, UP: 낮은값, LOW: 높은값 자동 기입
            d_mid = (z1 + z2) / 2
            d_min = min(z1, z2)
            d_max = max(z1, z2)
            
            self.paut_manual_vars['d'].set(f"{d_mid:.2f}")
            self.paut_manual_vars['D_Upper'].set(f"{d_min:.2f}")
            self.paut_manual_vars['D_Lower'].set(f"{d_max:.2f}")

            self._update_paut_auto_loc()
            self.log(f"📐 h 산출 완료 ({peak}% peak -> {db}dB 강하 목표 {target_fsh:.1f}%): {h_calc:.2f} mm (D:{d_mid:.2f}, U:{d_min:.2f}, L:{d_max:.2f})")
        except Exception as e:
            messagebox.showerror("계산 오류", f"입력값을 확인해주세요: {e}")

    def _calculate_paut_l(self):
        """[NEW] L1, L2 값을 이용한 결함 길이(L) 자동 계산"""
        try:
            l1 = float(self.paut_manual_vars['L1'].get() or 0)
            l2 = float(self.paut_manual_vars['L2'].get() or 0)
            l_calc = abs(l1 - l2)
            self.paut_manual_vars['l'].set(f"{l_calc:.2f}")
            self.log(f"📐 L 산출 완료 (L1:{l1} ~ L2:{l2}): {l_calc:.2f} mm")
        except Exception as e:
            messagebox.showerror("계산 오류", f"L1, L2 입력값을 확인해주세요: {e}")

    def _save_paut_manual_to_tree(self):
        """[NEW] 수동 판정 구역의 입력값을 선택된 트리뷰 행(들)에 저장"""
        selected_ids = self.paut_preview_tree.selection()
        if not selected_ids:
            messagebox.showwarning("선택 누락", "저장할 행을 먼저 선택해주세요.")
            return

        try:
            t = self.to_float(self.paut_manual_vars['t'].get() or 0)
            h = self.to_float(self.paut_manual_vars['h'].get() or 0)
            l = self.to_float(self.paut_manual_vars['l'].get() or 0)
            d = self.to_float(self.paut_manual_vars['d'].get() or 0)
            l1 = self.paut_manual_vars['L1'].get()
            l2 = self.paut_manual_vars['L2'].get()
            up = self.paut_manual_vars['D_Upper'].get()
            low = self.paut_manual_vars['D_Lower'].get()
            nat = self.paut_manual_vars['nature'].get()
            rem = "" # Remarks can be added to UI if needed, for now keep existing or clear

            # Run evaluation logic to get the result
            eval_mode = self.paut_eval_mode.get()
            res, loc = self.evaluate_paut_flaw(t, h, l, d, nat, mode=eval_mode)
            
            al = f"{h/l:.3f}" if l > 0 else ""
            at = f"{h/t:.3f}" if t > 0 else ""

            mode_info = self._get_mode_info("PAUT")
            if not mode_info: return
            _, idx_map, data_list, _ = mode_info

            for item_id in selected_ids:
                view_idx = self.paut_preview_tree.index(item_id)
                if 0 <= view_idx < len(idx_map):
                    actual_idx = idx_map[view_idx]
                    item = data_list[actual_idx]
                    item["Th'k(mm)"] = t
                    item['Height(mm)'] = h
                    item['Length(mm)'] = l
                    item['Start'] = l1
                    item['End'] = l2
                    item['Upper'] = up
                    item['Lower'] = low
                    item['Type of Flaw'] = nat
                    item['Evaluation'] = f"{res} ({loc})"
                    item['a/l'] = al
                    item['a/t'] = at
            
            self.populate_preview(data_list, mode="PAUT", switch_tab=False)
            self.log(f"💾 선택된 {len(selected_ids)}개 행에 수동 판정 결과 저장 완료.")
            
            # [NEW] 저장 후 입력 필드 초기화
            self._reset_paut_manual_inputs()
            
        except Exception as e:
            messagebox.showerror("저장 오류", f"입력값을 확인해주세요: {e}")

    def _run_manual_paut_eval(self):
        try:
            t = float(self.paut_manual_vars['t'].get() or 0)
            h = float(self.paut_manual_vars['h'].get() or 0)
            l = float(self.paut_manual_vars['l'].get() or 0)
            d = float(self.paut_manual_vars['d'].get() or 0)
            nat = self.paut_manual_vars['nature'].get()
            eval_mode = self.paut_eval_mode.get()
            
            res, loc = self.evaluate_paut_flaw(t, h, l, d, nat, mode=eval_mode)
            self.paut_manual_vars['loc'].set(loc)
            if res == "Accept":
                self.paut_res_label.config(text=f"{res} (위치: {loc})", fg="white", bg="#27ae60")
            else:
                self.paut_res_label.config(text=f"{res} (위치: {loc})", fg="white", bg="#e74c3c")
        except Exception as e:
            messagebox.showerror("입력 오류", f"입력값을 확인해주세요: {e}")

    def _extract_paut_data(self):
        file_path = self.paut_target_file_path.get()
        if not file_path:
            messagebox.showwarning("파일 미선택", "PAUT 데이터 엑셀 파일을 먼저 선택해주세요.")
            return
            
        try:
            self.log(f"📂 PAUT 데이터 불러오는 중: {os.path.basename(file_path)}")
            self.progress['value'] = 20
            
            df = pd.read_excel(file_path)
            cols = df.columns.tolist()
            
            mapping = {
                "t": next((c for c in cols if any(x in c.upper() for x in ["THK", "두께", "TH'K", "THICKNESS"])), None),
                "h": next((c for c in cols if any(x in c.upper() for x in ["HEIGHT", "높이", "FLAW_HEIGHT"])), None),
                "l": next((c for c in cols if any(x in c.upper() for x in ["LENGTH", "길이", "FLAW_LENGTH", "LENTH"])), None),
                "d": next((c for c in cols if any(x in c.upper() for x in ["DEPTH", "깊이", "FLAW_DEPTH"])), None),
                "nature": next((c for c in cols if any(x in c.upper() for x in ["NAT", "종류", "FLAW_NATURE", "TYPE"])), None),
                "line": next((c for c in cols if any(x in c.upper() for x in ["LINE", "ISO", "DWG", "DRAWING"])), None),
                "joint": next((c for c in cols if any(x in c.upper() for x in ["JOINT", "WELD"])), None),
                "start": next((c for c in cols if any(x in c.upper() for x in ["START", "L1"])), None),
                "end": next((c for c in cols if any(x in c.upper() for x in ["END", "L2"])), None),
                "upper": next((c for c in cols if any(x in c.upper() for x in ["UPPER", "UP", "DEPTH"])), None),
                "lower": next((c for c in cols if any(x in c.upper() for x in ["LOWER", "LOW"])), None),
                "remarks": next((c for c in cols if any(x in c.upper() for x in ["REMARK", "REMAKER", "비고"])), None)
            }
            
            if not all([mapping["t"], mapping["h"], mapping["l"], mapping["d"]]):
                self.log("⚠️ 필수 컬럼을 모두 찾을 수 없어 자동 매핑에 실패했습니다.")
                
            self.paut_extracted_data = []
            for _, row in df.iterrows():
                item = {
                    'selected': True,
                    'Line No.': str(row.get(mapping["line"], "")) if mapping["line"] else "",
                    'Joint No.': str(row.get(mapping["joint"], "")) if mapping["joint"] else "",
                    'Th\'k(mm)': row.get(mapping["t"], 0),
                    'Start': row.get(mapping["start"], ""),
                    'End': row.get(mapping["end"], ""),
                    'Length(mm)': row.get(mapping["l"], 0),
                    'Upper': row.get(mapping["upper"], ""),
                    'Lower': row.get(mapping["lower"], ""),
                    'Height(mm)': row.get(mapping["h"], 0),
                    'Type of Flaw': str(row.get(mapping["nature"], "Slag")) if mapping["nature"] else "Slag",
                    'a/l': "", 'a/t': "",
                    'Evaluation': "",
                    'Remarks': str(row.get(mapping["remarks"], "")) if mapping["remarks"] else ""
                }
                # [NEW] 원본 엑셀의 모든 컬럼 데이터를 유지하여 '컬럼 관리'에서 활용 가능하게 함
                item_full = row.to_dict()
                item_full.update(item)
                # [NEW] 수동 추가된 컬럼(Remark 등)이 추출 시 증발하지 않도록 초기화
                for k in self.paut_column_keys:
                    if k not in item_full and k != "selected":
                        item_full[k] = ""
                self.paut_extracted_data.append(item_full)
            
            self.progress['value'] = 100
            self.sort_by_column("ISO/DWG", mode="PAUT") # Auto sort for PAUT
            self.log(f"✅ PAUT 데이터 추출 완료: {len(self.paut_extracted_data)} 건")
            
        except Exception as e:
            self.log(f"❌ PAUT 데이터 추출 오류: {e}")

    def _run_batch_paut_eval(self):
        if not self.paut_extracted_data:
            messagebox.showwarning("데이터 없음", "먼저 데이터를 추출해주세요.")
            return
            
        self.log("🚀 PAUT 일괄 판정 시작...")
        count_ok, count_ng = 0, 0
        
        for item in self.paut_extracted_data:
            t = self.to_float(item.get('Th\'k(mm)', 0))
            h = self.to_float(item.get('Height(mm)', 0))
            l = self.to_float(item.get('Length(mm)', 0))
            d = self.to_float(item.get('Upper', 0))
            nat = item.get('Type of Flaw', 'Slag')
            eval_mode = self.paut_eval_mode.get()
            
            res, loc = self.evaluate_paut_flaw(t, h, l, d, nat, mode=eval_mode)
            item['Evaluation'] = f"{res} ({loc})"
            # a/l, a/t calculation
            if l > 0: item['a/l'] = f"{h/l:.3f}"
            if t > 0: item['a/t'] = f"{h/t:.3f}"
            
            if res == "Accept": count_ok += 1
            else: count_ng += 1
            
        self.populate_preview(self.paut_extracted_data, mode="PAUT")
        self.log(f"✅ 판정 완료: 총 {len(self.paut_extracted_data)} 건 (합격: {count_ok}, 불합격: {count_ng})")

    def _generate_paut_report(self):
        self.save_settings() # Ensure UI -> config sync
        template_path = self.paut_template_file_path.get()
        if not template_path or not os.path.exists(template_path):
            messagebox.showwarning("파일 미선택", "PAUT 성적서 양식 파일을 선택해주세요.")
            return

        final_list = [d for d in self.paut_extracted_data if d.get('selected', True)]
        if not final_list:
            messagebox.showwarning("항목 미선택", "선택된 데이터가 없습니다.")
            return

        self.log(f"🚀 PAUT 성적서 생성 시작 (총 {len(final_list)} 건)...")
        self.progress['value'] = 0
        
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.worksheets[0]
            ws.title = f"PAUT_Report_001"
            
            self.add_logos_to_sheet(ws, is_cover=False)
            self.force_print_settings(ws, context="DATA")
            
            start_row = int(self.config.get('PAUT_START_ROW', 11))
            end_row = int(self.config.get('PAUT_DATA_END_ROW', 40))
            
            curr_row = start_row
            data_font = Font(size=9)
            
            for i, item in enumerate(final_list):
                if curr_row > end_row:
                    ws = self.prepare_next_paut_sheet(wb, ws.title, i//(end_row-start_row+1))
                    curr_row = start_row
                
                mapping = {
                    self.config.get('PAUT_COL_NO', '1'): item.get('No', i+1), 
                    self.config.get('PAUT_COL_LINE', '2'): item.get('Line No.', ''), 
                    self.config.get('PAUT_COL_JOINT', '3'): item.get('Joint No.', ''),
                    self.config.get('PAUT_COL_THK', '4'): item.get('Th\'k(mm)', ''), 
                    self.config.get('PAUT_COL_H', '5'): item.get('Height(mm)', ''), 
                    self.config.get('PAUT_COL_L', '6'): item.get('Length(mm)', ''),
                    self.config.get('PAUT_COL_UP', '7'): item.get('Upper', ''), 
                    self.config.get('PAUT_COL_LOW', '0'): item.get('Lower', ''),
                    self.config.get('PAUT_COL_START', '0'): item.get('Start', ''),
                    self.config.get('PAUT_COL_END', '0'): item.get('End', ''),
                    self.config.get('PAUT_COL_AL', '0'): item.get('a/l', ''),
                    self.config.get('PAUT_COL_AT', '0'): item.get('a/t', ''),
                    self.config.get('PAUT_COL_NAT', '8'): item.get('Type of Flaw', ''), 
                    self.config.get('PAUT_COL_EVAL', '9'): item.get('Evaluation', ''),
                    self.config.get('PAUT_COL_REM', '10'): item.get('Remarks', '')
                }
                
                for col_key, val in mapping.items():
                    try:
                        col_idx = self.col_to_num(col_key)
                        if col_idx < 1: continue # Skip if 0 or invalid
                        
                        cell = ws.cell(row=curr_row, column=col_idx, value=val)
                        cell.font = data_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    except Exception:
                        continue
                
                curr_row += 1
                self.progress['value'] = (i+1) / len(final_list) * 100
            
            out_name = f"PAUT_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            out_path = os.path.join(BASE_DIR, "output", out_name) if os.path.exists(os.path.join(BASE_DIR, "output")) else os.path.join(BASE_DIR, out_name)
            
            if not os.path.exists(os.path.dirname(out_path)): os.makedirs(os.path.dirname(out_path))
            wb.save(out_path)
            self.log(f"✅ PAUT 성적서 생성 완료: {out_name}")
            os.startfile(os.path.dirname(out_path))
            
        except Exception as e:
            self.log(f"❌ PAUT 성적서 생성 오류: {e}")

    def _export_paut_session(self):
        """[NEW] 현재 PAUT 미리보기 데이터를 CSV로 내보내기"""
        if not self.paut_extracted_data:
            messagebox.showwarning("데이터 없음", "내보낼 데이터가 없습니다.")
            return
        
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="PAUT 세션 저장",
            initialfile=f"PAUT_Session_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        if not path: return

        try:
            df = pd.DataFrame(self.paut_extracted_data)
            # UTF-8-SIG for Excel compatibility
            df.to_csv(path, index=False, encoding='utf-8-sig')
            self.log(f"💾 PAUT 세션 저장 완료: {os.path.basename(path)}")
            messagebox.showinfo("저장 완료", f"현재 세션이 저장되었습니다.\n{os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("저장 오류", f"세션 저장 중 오류 발생: {e}")

    def _import_paut_session(self):
        """[NEW] 저장된 PAUT 세션 CSV 불러오기"""
        path = filedialog.askopenfilename(
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="PAUT 세션 불러오기"
        )
        if not path: return

        try:
            df = pd.read_csv(path)
            # Ensure mandatory columns or basic structure
            new_data = df.to_dict('records')
            
            # Simple normalization (handle NaNs from CSV)
            normalized_data = []
            for d in new_data:
                clean_dict = {}
                for k, v in d.items():
                    if pd.isna(v): clean_dict[k] = ""
                    else: clean_dict[k] = v
                normalized_data.append(clean_dict)
            
            self.paut_extracted_data = normalized_data
            self.file_info_vars['PAUT'].set(f"📂 로드됨: {os.path.basename(path)} (총 {len(normalized_data)}건)")
            self.populate_preview(self.paut_extracted_data, mode="PAUT")
            self.log(f"📂 PAUT 세션 불러오기 완료: {os.path.basename(path)} ({len(normalized_data)}건)")
        except Exception as e:
            messagebox.showerror("불러오기 오류", f"세션 불러오기 중 오류 발생: {e}")

    def _reset_paut_manual_inputs(self):
        """[NEW] 개별 판정 입력 필드 초기화 (두께 T 등 공통 설정값 제외)"""
        for k in ['h', 'l', 'd', 'z1', 'z2', 'L1', 'L2', 'D_Upper', 'D_Lower']:
            if k in self.paut_manual_vars:
                self.paut_manual_vars[k].set("")
        self.paut_manual_vars['nature'].set("Slag")
        self.paut_manual_vars['loc'].set("-")
        self.paut_res_label.config(text="결과 대기", fg="black", bg="#f3f4f6")

    def prepare_next_paut_sheet(self, wb, prev_title, page_num):
        source_sheet = wb[prev_title]
        new_sheet = wb.copy_worksheet(source_sheet)
        new_sheet.title = f"PAUT_Report_{page_num+1:03d}"
        
        start_row = int(self.config.get('PAUT_START_ROW', 11))
        end_row = int(self.config.get('PAUT_DATA_END_ROW', 40))
        for r in range(start_row, end_row + 1):
            for c in range(1, 11): new_sheet.cell(row=r, column=c).value = None
        
        self.add_logos_to_sheet(new_sheet, is_cover=False)
        return new_sheet

    def col_to_num(self, col_str):
        """[NEW] 엑셀 열 이름(A, B, AA...)을 숫자 인덱스(1, 2, 27...)로 변환"""
        if not col_str: return 0
        s = str(col_str).strip().upper().replace("열", "").replace("행", "").strip()
        if s.isdigit(): return int(s)
        
        res = 0
        for i, char in enumerate(reversed(s)):
            if ord('A') <= ord(char) <= ord('Z'):
                res += (ord(char) - ord('A') + 1) * (26 ** i)
            else:
                return 0 # Invalid
        return res

    def _create_tree_context_menu(self):
        """미리보기 트리뷰용 공통 컨텍스트 메뉴 초기화"""
        self.ctx_menu = tk.Menu(self.root, tearoff=0)
        # 기본 항목 기입 (show_context_menu에서 command 동적 바인딩)
        self.ctx_menu.add_command(label="선택 항목 체크/해제 토글 (Toggle Check)")
        self.ctx_menu.add_command(label="현재 셀 내용만 복사 (Copy Cell)")
        self.ctx_menu.add_command(label="선택 행 전체 복사 (Copy Row)")
        self.ctx_menu.add_command(label="이 컬럼 전체 복사 (Copy Column)")
        self.ctx_menu.add_command(label="셀 내용 붙여넣기 (Paste)")
        self.ctx_menu.add_separator()
        self.ctx_menu.add_command(label="위로 이동 (Move Up)")
        self.ctx_menu.add_command(label="아래로 이동 (Move Down)")
        self.ctx_menu.add_separator()
        self.ctx_menu.add_command(label="행 추가 (Add Row)")
        self.ctx_menu.add_command(label="선택 삭제 (Delete Selected)")
        self.ctx_menu.add_separator()
        self.ctx_menu.add_command(label="선택 항목 ISO 병합 (Merge ISO)")
        self.ctx_menu.add_command(label="선택 항목 Joint 병합 (Merge Joint)")
        self.ctx_menu.add_command(label="선택 항목 ISO 병합 해제 (Ungroup ISO)")
        self.ctx_menu.add_command(label="선택 항목 Joint 병합 해제 (Ungroup Joint)")
        self.ctx_menu.add_separator()
        self.ctx_menu.add_command(label="선택 항목 일괄 변경 (Bulk Update)")
        # PAUT 전용 (기본은 숨김)
        self.ctx_menu.add_separator()
        self.ctx_menu.add_command(label="컬럼 설정 (Column Manager)")

    def _create_entry_context_menu(self):
        self.entry_popup = tk.Menu(self.root, tearoff=0)
        self.entry_popup.add_command(label="복사 (Copy)", command=self._copy_entry)
        self.entry_popup.add_command(label="붙여넣기 (Paste)", command=self._paste_entry)
        self.entry_popup.add_separator()
        self.entry_popup.add_command(label="잘라내기 (Cut)", command=self._cut_entry)
        
        self.root.bind_class("Entry", "<Button-3>", self._show_entry_popup)
        self.root.bind_class("TEntry", "<Button-3>", self._show_entry_popup)

    def _show_entry_popup(self, event):
        self.current_entry = event.widget
        self._popup_selection_range = None
        if self.current_entry.select_present():
            self._popup_selection_range = (self.current_entry.index("sel.first"), self.current_entry.index("sel.last"))
        self.entry_popup.tk_popup(event.x_root, event.y_root)

    def _copy_entry(self):
        try:
            if getattr(self, 'current_entry', None):
                if self.current_entry.select_present():
                    first, last = self.current_entry.index("sel.first"), self.current_entry.index("sel.last")
                elif getattr(self, '_popup_selection_range', None):
                    first, last = self._popup_selection_range
                else:
                    return
                text = self.current_entry.get()[first:last]
                self.root.clipboard_clear()
                self.root.clipboard_append(text)
                self.root.update()
        except Exception: pass

    def _paste_entry(self):
        try:
            text = self.root.clipboard_get()
            if getattr(self, 'current_entry', None):
                if getattr(self, '_popup_selection_range', None):
                    first, last = self._popup_selection_range
                    self.current_entry.delete(first, last)
                    self.current_entry.insert(first, text)
                elif self.current_entry.select_present():
                    self.current_entry.delete("sel.first", "sel.last")
                    self.current_entry.insert(tk.INSERT, text)
                else:
                    self.current_entry.insert(tk.INSERT, text)
        except Exception: pass

    def _cut_entry(self):
        self._copy_entry()
        try:
            if getattr(self, 'current_entry', None):
                if getattr(self, '_popup_selection_range', None):
                    first, last = self._popup_selection_range
                    self.current_entry.delete(first, last)
                elif self.current_entry.select_present():
                    self.current_entry.delete("sel.first", "sel.last")
        except Exception: pass

    def _create_gapji_preview_ui(self, parent, mode):
        """Create the Gapji (Cover) preview tab content."""
        container = tk.Frame(parent, background="#f3f4f6")
        container.pack(fill='both', expand=True)
        
        # Toolbar
        toolbar = tk.Frame(container, background="#ffffff", pady=5, padx=10, highlightthickness=1, highlightbackground="#e5e7eb")
        toolbar.pack(fill='x')
        ttk.Label(toolbar, text=f"📄 {mode} 갑지 레이아웃 미리보기", font=("Malgun Gothic", 9, "bold"), background="#ffffff").pack(side='left')
        
        btn_f = tk.Frame(toolbar, background="#ffffff")
        btn_f.pack(side='right')
        ttk.Button(btn_f, text=" 🚀 이대로 성적서 생성 ", command=self.run_process).pack(side='right')
        ttk.Button(btn_f, text=" 🔄 새로고침 ", command=lambda: self._update_gapji_preview(mode)).pack(side='right', padx=5)
        
        # Scrollable area
        outer_f = tk.Frame(container, background="#9ca3af") # Desk background
        outer_f.pack(fill='both', expand=True)
        
        # Canvas
        self.gapji_canvases = getattr(self, 'gapji_canvases', {})
        # A4 @ 72dpi is ~595x842. We'll use a slightly smaller view to fit.
        self.gapji_scale = 2.5 # 1mm = 2.5px
        canv_w, canv_h = int(210 * self.gapji_scale), int(297 * self.gapji_scale)
        
        canv = tk.Canvas(outer_f, background="#ffffff", width=canv_w, height=canv_h, 
                         highlightthickness=1, highlightbackground="#374151")
        canv.pack(pady=20)
        self.gapji_canvases[mode] = canv

    def _excel_anchor_to_px(self, anchor, scale):
        """Roughly map Excel cell to A4 pixel coordinates."""
        import re
        m = re.match(r"([A-Z]+)(\d+)", anchor.upper())
        if not m: return 0, 0
        
        col_str, row_str = m.groups()
        # Col to Num (A=0, B=1...)
        col = 0
        for char in col_str: col = col * 26 + (ord(char) - ord('A'))
        row = int(row_str) - 1
        
        # Rough estimates for A4 grid (210mm / 10 cols, 297mm / 50 rows)
        # Adjusting to feel more like typical report templates
        x = (col * 15) * scale # 15mm per col
        y = (row * 6) * scale  # 6mm per row
        return x, y

    def _update_gapji_preview(self, mode):
        """Render a visual representation of the Gapji on the Canvas."""
        if not hasattr(self, 'gapji_canvases') or mode not in self.gapji_canvases: return
        canv = self.gapji_canvases[mode]
        canv.delete("all")
        
        scale = getattr(self, 'gapji_scale', 2.5)
        
        # 1. Draw Title Text (Approximation of Template)
        canv.create_text(105*scale, 40*scale, text=f"{mode} EXAMINATION REPORT", 
                         font=("Arial", 16, "bold"), fill="#1e3a8a")
        canv.create_line(30*scale, 50*scale, 180*scale, 50*scale, fill="#1e3a8a", width=2)
        
        # 2. Draw Metadata
        y_start = 120 # mm
        info = [
            ("PROJECT:", self.gapji_project.get(), y_start),
            ("CUSTOMER:", self.gapji_customer.get(), y_start + 12),
            ("ITEM:", self.gapji_item.get(), y_start + 24),
            ("MATERIAL:", self.gapji_material.get(), y_start + 36),
            ("DATE:", self.gapji_exam_date.get(), y_start + 48),
            ("REPORT NO:", self.gapji_report_no.get(), y_start + 60),
        ]
        
        for label, val, y in info:
            canv.create_text(40*scale, y*scale, text=label, font=("Arial", 9, "bold"), anchor='w')
            canv.create_text(80*scale, y*scale, text=val if val else "(입력 안 됨)", 
                             font=("Arial", 9), anchor='w', fill="#374151" if val else "#9ca3af")
            canv.create_line(80*scale, (y+2)*scale, 180*scale, (y+2)*scale, fill="#e5e7eb")

        # 3. Draw Logos
        from PIL import Image, ImageTk
        logo_types = [
            ("SITCO", ["로고", "SITCO"]),
            ("SEOUL", ["서울검사", "SEOUL"]),
            ("FOOTER", ["바닥글", "FOOTER"]),
            ("FOOTER_PT", ["바닥글", "PT", "LOGO"])
        ]
        
        if not hasattr(self, '_preview_img_refs'): self._preview_img_refs = {}
        
        for lt, keywords in logo_types:
            prefix = f"{lt}_{mode}_COVER"
            if prefix not in self.config and f"{lt}_COVER" in self.config: prefix = f"{lt}_COVER"
            
            path = self.config.get(f"{prefix}_PATH", "")
            if not path:
                # Fallback to smart search if empty
                found = self.find_image_smart(keywords[0])
                if found: path = found
            
            if path and os.path.exists(path):
                try:
                    img = Image.open(path)
                    w_mm = float(self.config.get(f"{prefix}_W", 40))
                    h_mm = float(self.config.get(f"{prefix}_H", 15))
                    x_off = float(self.config.get(f"{prefix}_X", 0))
                    y_off = float(self.config.get(f"{prefix}_Y", 0))
                    anchor = self.config.get(f"{prefix}_ANCHOR", "A1")
                    
                    # Convert to pixels
                    w_px, h_px = int(w_mm * scale), int(h_mm * scale)
                    ax, ay = self._excel_anchor_to_px(anchor, scale)
                    
                    final_x = ax + (x_off * scale)
                    final_y = ay + (y_off * scale)
                    
                    # Scale and keep reference
                    img = img.resize((max(1, w_px), max(1, h_px)), Image.Resampling.LANCZOS)
                    photo = ImageTk.PhotoImage(img)
                    self._preview_img_refs[f"{mode}_{lt}"] = photo
                    
                    canv.create_image(final_x, final_y, image=photo, anchor='nw')
                except Exception as e:
                    print(f"Preview error for {lt}: {e}")

    def _create_gapji_meta_ui(self, parent, use_pack=True):
        block = tk.LabelFrame(parent, text=" 리포트 기본 정보 (Gapji Info) ", padx=10, pady=5, background="#ffffff", font=("Malgun Gothic", 9, "bold"))
        if use_pack:
            block.pack(fill='x', pady=5, padx=2)
        else:
            block.grid(row=0, column=0, sticky='ew', pady=5, padx=2)
        
        # Grid for info entries
        for col in [1, 3]: block.columnconfigure(col, weight=1)
        
        fields = [
            ("공사명:", self.gapji_project, 0, 0),
            ("발주처:", self.gapji_customer, 0, 2),
            ("품명:", self.gapji_item, 1, 0),
            ("재질:", self.gapji_material, 1, 2),
            ("리포트번호:", self.gapji_report_no, 2, 0),
            ("검사일자:", self.gapji_exam_date, 2, 2)
        ]
        
        for lbl, var, r, c in fields:
            tk.Label(block, text=lbl, background="#ffffff", font=("Malgun Gothic", 8)).grid(row=r, column=c, sticky='e', padx=2, pady=2)
            ttk.Entry(block, textvariable=var, width=15).grid(row=r, column=c+1, sticky='ew', padx=2, pady=2)
            # Add to setting_vars for auto-save
            cfg_key = f"GAPJI_{lbl.replace(':', '').upper()}"
            if "공사명" in lbl: cfg_key = "GAPJI_PROJECT"
            elif "발주처" in lbl: cfg_key = "GAPJI_CUSTOMER"
            elif "품명" in lbl: cfg_key = "GAPJI_ITEM"
            elif "재질" in lbl: cfg_key = "GAPJI_MATERIAL"
            elif "리포트번호" in lbl: cfg_key = "GAPJI_REPORT_NO"
            elif "검사일자" in lbl: cfg_key = "GAPJI_EXAM_DATE"
            self.setting_vars[cfg_key] = var
            
        # Add a refresh button for preview
        btn_f = tk.Frame(block, background="#ffffff")
        btn_f.grid(row=3, column=0, columnspan=4, pady=5, sticky='ew')
        ttk.Button(btn_f, text="✨ 갑지 미리보기 업데이트", command=self._update_gapji_preview_current).pack(side='top', fill='x', padx=5)

    def _update_gapji_preview_current(self, event=None):
        """통합된 RT 탭 및 서브 미리보기 구조에 맞춰 갑지 미리보기 갱신"""
        try:
            tab_text = self.mode_notebook.tab(self.mode_notebook.select(), "text")
            mode = "PMI"
            nb_inner = None
            
            if "RT" in tab_text:
                mode = self.rt_sub_mode.get()
                nb_inner = getattr(self, 'rt_preview_nb', None)
            elif "PT" in tab_text:
                mode = "PT"
                # PT는 기존 preview_notebooks 구조 사용 가능 여부 확인
                if hasattr(self, 'preview_notebooks'): nb_inner = self.preview_notebooks.get("PT")
            elif "PAUT" in tab_text:
                mode = "PAUT"
                if hasattr(self, 'preview_notebooks'): nb_inner = self.preview_notebooks.get("PAUT")
            elif "PMI" in tab_text or "OES" in tab_text:
                mode = "PMI"
                if hasattr(self, 'preview_notebooks'): nb_inner = self.preview_notebooks.get("PMI")

            # 2. Logic for tab change event
            if event and event.widget:
                try:
                    nb = event.widget
                    inner_tab_text = nb.tab(nb.select(), "text")
                    if "갑지" in inner_tab_text:
                        if nb_inner: nb_inner.select(2) # [Standard, KOGAS, Gapji] -> Gapji is 2
                        self._update_gapji_preview(mode)
                    elif "미리보기" in inner_tab_text:
                        if nb_inner:
                            # 만약 서브 미리보기가 가스공사라면 1번, 아니면 0번
                            target_idx = 1 if "가스공사" in inner_tab_text else 0
                            nb_inner.select(target_idx)
                except: pass
            else:
                # Manual trigger (Button)
                if nb_inner: nb_inner.select(2) # Force select Gapji tab
                self._update_gapji_preview(mode)
        except Exception as e:
            self.log(f"미리보기 업데이트 실패: {e}")

    def _create_setting_grid(self, parent, context):
        # [REFINED] More organized layout for logo settings
        items = [
            ("SITCO 로고", f"SITCO_{context}"), 
            ("서울검사 로고", f"SEOUL_{context}"), 
            ("바닥글 우측 (Footer Right)", f"FOOTER_{context}"), 
            ("바닥글 좌측 (Footer Left)", f"FOOTER_PT_{context}")
        ]
        
        next_row = 0
        for i, (label, key_prefix) in enumerate(items):
            # [HYPER-REACTIVE] Force columnspan=1 to match weighted col 0 of parent
            short_label = label.split(" ")[0].replace("로고", "")
            block = tk.LabelFrame(parent, text=f" {short_label} ", padx=2, pady=1, background="#ffffff", font=("Malgun Gothic", 8, "bold"))
            # [FIX] Offset row by 1 to leave room for Gapji Info at row 0 if present
            block.grid(row=i + 1, column=0, sticky='ew', pady=1, padx=2)
            
            frame = tk.Frame(block, background="#ffffff")
            frame.pack(fill='x')
            frame.columnconfigure(1, weight=5, minsize=0) # Path entry weight
            for c in [4, 6, 8, 10]: frame.columnconfigure(c, weight=1, minsize=0) # Coords weight
            
            tk.Label(frame, text="P:", width=2, anchor='e', background="#ffffff", font=("Malgun Gothic", 7)).grid(row=0, column=0)
            v_path = tk.StringVar(value=self.config.get(f"{key_prefix}_PATH", ""))
            ttk.Entry(frame, textvariable=v_path, width=15, exportselection=False).grid(row=0, column=1, sticky='ew', padx=1) # [FIX] Width 1 -> 15
            ttk.Button(frame, text="..", width=2, command=lambda v=v_path: self._browse_file(v, [("Images", "*.png;*.jpg;*.jpeg")])).grid(row=0, column=2, padx=1)
            self.setting_vars[f"{key_prefix}_PATH"] = v_path
            
            for idx, (coord, key_suffix) in enumerate([("X", "X"), ("Y", "Y"), ("W", "W"), ("H", "H")]):
                tk.Label(frame, text=f"{coord}:", width=1, anchor='e', background="#ffffff", font=("Malgun Gothic", 7)).grid(row=0, column=3+idx*2)
                v = tk.StringVar(value=str(self.config.get(f"{key_prefix}_{key_suffix}", "0.0")))
                ttk.Entry(frame, textvariable=v, width=6, exportselection=False).grid(row=0, column=4+idx*2, sticky='ew', padx=1) # [FIX] Width 1 -> 6
                self.setting_vars[f"{key_prefix}_{key_suffix}"] = v
            next_row = i + 1
        return next_row
            
    def _create_margin_settings(self, parent, context, use_pack=True):
        # [FIX] Adaptive layout to avoid TclError (grid vs pack conflict)
        frame = ttk.LabelFrame(parent, text=" 인쇄 및 여백 (Margins) ", padding=2)
        if use_pack:
            frame.pack(fill='x', pady=(5, 0))
        else:
            # Inside Notebook tabs, we usually use grid. 
            # We assume it's placed at the bottom of the existing grid.
            frame.grid(row=100, column=0, columnspan=10, sticky='ew', pady=(10, 0))
        
        # Force 0 minsize for elasticity
        for col in range(10): frame.columnconfigure(col, minsize=0)
        # [FIX] Distribute expansion only to entries
        for col in [0, 2, 4, 6, 8]: frame.columnconfigure(col, weight=0) # Labels
        for col in [1, 3, 5, 7, 9]: frame.columnconfigure(col, weight=1) # Entries

        m_items = [("T", "TOP"), ("B", "BOTTOM"), ("L", "LEFT"), ("R", "RIGHT")]
        for i, (sh_lbl, key) in enumerate(m_items):
            ttk.Label(frame, text=f"{sh_lbl}:", font=("Arial", 7)).grid(row=0, column=i*2, sticky='e', padx=1)
            v = tk.StringVar(value=str(self.config.get(f"MARGIN_{context}_{key}", "0.2")))
            ent = ttk.Entry(frame, textvariable=v, width=1)
            ent.grid(row=0, column=i*2+1, sticky='ew', padx=1)
            ent.bind("<Return>", lambda e: self.root.focus_set())
            self.setting_vars[f"MARGIN_{context}_{key}"] = v
            
        ttk.Label(frame, text="S:", font=("Arial", 7)).grid(row=0, column=8, sticky='e', padx=1)
        v_s = tk.StringVar(value=str(self.config.get(f"PRINT_SCALE_{context}", "95")))
        ent_s = ttk.Entry(frame, textvariable=v_s, width=1)
        ent_s.grid(row=0, column=9, sticky='ew', padx=1)
        ent_s.bind("<Return>", lambda e: self.root.focus_set())
        self.setting_vars[f"PRINT_SCALE_{context}"] = v_s

        # [NEW] Print Area Row
        area_f = ttk.Frame(frame)
        area_f.grid(row=1, column=0, columnspan=10, sticky='ew', pady=(2, 0))
        ttk.Label(area_f, text="Area:", font=("Arial", 7)).pack(side='left', padx=1)
        v_a = tk.StringVar(value=str(self.config.get(f"PRINT_AREA_{context}", "")))
        ent_a = ttk.Entry(area_f, textvariable=v_a)
        ent_a.pack(side='left', fill='x', expand=True, padx=1)
        self.setting_vars[f"PRINT_AREA_{context}"] = v_a
        
        # Adjusters Row
        sub = ttk.Frame(frame)
        sub.grid(row=2, column=0, columnspan=10, sticky='ew', pady=2)
        for c in [1, 3, 5, 7]: sub.columnconfigure(c, weight=1, minsize=0)
        for c in [0, 2, 4, 6]: sub.columnconfigure(c, weight=0, minsize=0)
        
        labels = [("Rows:", f"CUSTOM_ROWS_{context}"), ("H:", f"CUSTOM_ROW_HEIGHT_{context}"), 
                  ("Cols:", f"CUSTOM_COLS_{context}"), ("W:", f"CUSTOM_COL_WIDTH_{context}")]
        
        for i, (lbl, key) in enumerate(labels):
            ttk.Label(sub, text=lbl, font=("Arial", 7)).grid(row=0, column=i*2, sticky='e')
            defaultv = "16.5" if "HEIGHT" in key else ("10.0" if "WIDTH" in key else "")
            v = tk.StringVar(value=str(self.config.get(key, defaultv)))
            ent = ttk.Entry(sub, textvariable=v, width=1)
            ent.grid(row=0, column=i*2+1, sticky='ew', padx=1)
            ent.bind("<Return>", lambda e: self.root.focus_set())
            self.setting_vars[key] = v

    def _create_column_mapping_ui(self, parent, mode, mapping_items):
        """[NEW] 전 모듈 공용 엑셀 컬럼 매핑 및 이름 설정 UI (스크롤바 추가)"""
        # Outer Frame for Scrollbar
        outer_frame = tk.Frame(parent, background="#f9fafb")
        outer_frame.pack(fill='both', expand=True)
        
        canvas = tk.Canvas(outer_frame, background="#f9fafb", highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, background="#f9fafb")

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Mouse wheel support
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        container = ttk.LabelFrame(scrollable_frame, text=f" {mode} 성적서 및 미리보기 컬럼 설정 ", padding=10)
        container.pack(fill='both', expand=True, pady=5, padx=10)
        
        # Grid settings
        container.columnconfigure(2, weight=1)
        container.columnconfigure(5, weight=1)
        
        # Header labels
        ttk.Label(container, text="항목", font=("Malgun Gothic", 8, "bold")).grid(row=0, column=0, sticky='w')
        ttk.Label(container, text="엑셀열 | 표시이름", font=("Malgun Gothic", 8, "bold")).grid(row=0, column=1, columnspan=2, sticky='w')
        ttk.Label(container, text="항목", font=("Malgun Gothic", 8, "bold")).grid(row=0, column=3, sticky='w')
        ttk.Label(container, text="엑셀열 | 표시이름", font=("Malgun Gothic", 8, "bold")).grid(row=0, column=4, columnspan=2, sticky='w')
        
        last_row = 0
        for i, (label, key_idx, def_idx, key_name, def_name, internal_id) in enumerate(mapping_items):
            row = (i // 2) + 1
            last_row = max(last_row, row)
            col_offset = (i % 2) * 3
            
            # 1. Label
            ttk.Label(container, text=label, font=("Malgun Gothic", 8)).grid(row=row, column=col_offset, sticky='e', padx=2, pady=2)
            
            # 2. Excel Col #
            v_idx = tk.StringVar(value=str(self.config.get(key_idx, def_idx)))
            ent_idx = ttk.Entry(container, textvariable=v_idx, width=4)
            ent_idx.grid(row=row, column=col_offset+1, sticky='w', padx=2, pady=2)
            ent_idx.bind("<Return>", lambda e: self.root.focus_set())
            self.setting_vars[key_idx] = v_idx
            
            # 3. Display Name
            v_name = tk.StringVar(value=str(self.config.get(key_name, def_name)))
            ent_name = ttk.Entry(container, textvariable=v_name, width=12)
            ent_name.grid(row=row, column=col_offset+2, sticky='ew', padx=2, pady=2)
            ent_name.bind("<Return>", lambda e: self.root.focus_set())
            self.setting_vars[key_name] = v_name
            
            # Trace to update config and preview tree heading in real-time
            v_name.trace_add("write", lambda *args, tid=internal_id, m=mode, var=v_name, kn=key_name: self._update_mode_heading(m, tid, var, kn))

        ttk.Label(container, text="* 엑셀열: 알파벳(A, B...) 또는 숫자(1, 2...) 입력 (0=제외)", foreground="gray", font=("Malgun Gothic", 8)).grid(row=last_row+1, column=0, columnspan=6, pady=10)

    def _create_kogas_separated_column_mapping_ui(self, parent):
        """[NEW] 가스공사(KOGAS) 모드 전용 의뢰서(읽기) 및 보고서(쓰기) 분리 컬럼 설정 UI"""
        # Outer Frame for Scrollbar
        outer_frame = tk.Frame(parent, background="#f9fafb")
        outer_frame.pack(fill='both', expand=True)
        
        canvas = tk.Canvas(outer_frame, background="#f9fafb", highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, background="#f9fafb")

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Mouse wheel support
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Main horizontal container to hold Read and Write side-by-side
        main_container = tk.Frame(scrollable_frame, background="#f9fafb")
        main_container.pack(fill='both', expand=True, pady=5, padx=5)

        # 1. Left LabelFrame: [의뢰서 읽기] 설정
        read_frame = ttk.LabelFrame(main_container, text=" 📄 [의뢰서 읽기] 컬럼 설정 (Input) ", padding=10)
        read_frame.pack(side='left', fill='both', expand=True, padx=5, pady=5)

        # 2. Right LabelFrame: [보고서 쓰기] 설정
        write_frame = ttk.LabelFrame(main_container, text=" 📝 [보고서 쓰기] 컬럼 설정 (Output) ", padding=10)
        write_frame.pack(side='right', fill='both', expand=True, padx=5, pady=5)

        # Helper to render mappings inside a frame
        def render_fields(frame, items, is_read):
            frame.columnconfigure(2, weight=1)
            # Header
            ttk.Label(frame, text="항목", font=("Malgun Gothic", 8, "bold")).grid(row=0, column=0, sticky='w', pady=(0,5))
            ttk.Label(frame, text="열", font=("Malgun Gothic", 8, "bold")).grid(row=0, column=1, sticky='w', pady=(0,5))
            ttk.Label(frame, text="이름", font=("Malgun Gothic", 8, "bold")).grid(row=0, column=2, sticky='w', pady=(0,5))

            for idx, (label, key_idx, def_idx, key_name, def_name, internal_id) in enumerate(items):
                row = idx + 1
                ttk.Label(frame, text=label, font=("Malgun Gothic", 8)).grid(row=row, column=0, sticky='w', padx=2, pady=2)
                
                v_idx = tk.StringVar(value=str(self.config.get(key_idx, def_idx)))
                ent_idx = ttk.Entry(frame, textvariable=v_idx, width=4)
                ent_idx.grid(row=row, column=1, sticky='w', padx=2, pady=2)
                ent_idx.bind("<Return>", lambda e: self.root.focus_set())
                self.setting_vars[key_idx] = v_idx

                v_name = tk.StringVar(value=str(self.config.get(key_name, def_name)))
                ent_name = ttk.Entry(frame, textvariable=v_name, width=12)
                ent_name.grid(row=row, column=2, sticky='ew', padx=2, pady=2)
                ent_name.bind("<Return>", lambda e: self.root.focus_set())
                self.setting_vars[key_name] = v_name

                # Real-time heading update (only update for tree preview if it's the Read setting since tree is built on Read data!)
                if is_read:
                    v_name.trace_add("write", lambda *args, tid=internal_id, var=v_name, kn=key_name: self._update_mode_heading("KOGAS", tid, var, kn))
                else:
                    v_name.trace_add("write", lambda *args, kn=key_name, var=v_name: self.config.update({kn: var.get()}))

        # Read items
        read_items = [
            ("No:", "KOGAS_R_COL_NO", "1", "KOGAS_R_NAME_NO", "No", "No"),
            ("Date:", "KOGAS_R_COL_DATE", "2", "KOGAS_R_NAME_DATE", "Date", "Date"),
            ("Dwg No.:", "KOGAS_R_COL_DWG", "3", "KOGAS_R_NAME_DWG", "Dwg No", "Dwg"),
            ("Joint No.:", "KOGAS_R_COL_JOINT", "4", "KOGAS_R_NAME_JOINT", "Joint No", "Joint"),
            ("Location:", "KOGAS_R_COL_LOC", "5", "KOGAS_R_NAME_LOC", "Location", "Loc"),
            ("T:", "KOGAS_R_COL_THK", "6", "KOGAS_R_NAME_THK", "T", "T"),
            ("Mat:", "KOGAS_R_COL_MAT", "7", "KOGAS_R_NAME_MAT", "Mat", "Mat"),
            ("구경(Size):", "KOGAS_R_COL_SIZE", "8", "KOGAS_R_NAME_SIZE", "Size", "Size"),
            ("판정(Result):", "KOGAS_R_COL_RES", "9", "KOGAS_R_NAME_RES", "Result", "Result"),
            ("용접사:", "KOGAS_R_COL_WELDER", "10", "KOGAS_R_NAME_WELDER", "Welder", "Welder"),
            ("비고:", "KOGAS_R_COL_REM", "11", "KOGAS_R_NAME_REM", "Remarks", "Remarks"),
            ("① Crack:", "KOGAS_R_COL_D1", "12", "KOGAS_R_NAME_D1", "① Crack", "D1"),
            ("② IP:", "KOGAS_R_COL_D2", "13", "KOGAS_R_NAME_D2", "② IP", "D2"),
            ("③ LF:", "KOGAS_R_COL_D3", "14", "KOGAS_R_NAME_D3", "③ LF", "D3"),
            ("④ Slag:", "KOGAS_R_COL_D4", "15", "KOGAS_R_NAME_D4", "④ Slag", "D4"),
            ("⑤ Por:", "KOGAS_R_COL_D5", "16", "KOGAS_R_NAME_D5", "⑤ Por", "D5"),
            ("⑥ U/C:", "KOGAS_R_COL_D6", "17", "KOGAS_R_NAME_D6", "⑥ U/C", "D6"),
            ("⑦ RUC:", "KOGAS_R_COL_D7", "18", "KOGAS_R_NAME_D7", "⑦ RUC", "D7"),
            ("⑧ BT:", "KOGAS_R_COL_D8", "19", "KOGAS_R_NAME_D8", "⑧ BT", "D8"),
            ("⑨ IC:", "KOGAS_R_COL_D9", "20", "KOGAS_R_NAME_D9", "⑨ IC", "D9"),
            ("⑩ CP:", "KOGAS_R_COL_D10", "21", "KOGAS_R_NAME_D10", "⑩ CP", "D10"),
            ("⑪ RC:", "KOGAS_R_COL_D11", "22", "KOGAS_R_NAME_D11", "⑪ RC", "D11"),
            ("⑫ Mis:", "KOGAS_R_COL_D12", "23", "KOGAS_R_NAME_D12", "⑫ Mis", "D12"),
            ("⑬ EP:", "KOGAS_R_COL_D13", "24", "KOGAS_R_NAME_D13", "⑬ EP", "D13"),
            ("⑭ SD:", "KOGAS_R_COL_D14", "25", "KOGAS_R_NAME_D14", "⑭ SD", "D14"),
            ("⑮ Oth:", "KOGAS_R_COL_D15", "26", "KOGAS_R_NAME_D15", "⑮ Oth", "D15")
        ]

        # Write items
        write_items = [
            ("No:", "KOGAS_W_COL_NO", "1", "KOGAS_W_NAME_NO", "No", "No"),
            ("Date:", "KOGAS_W_COL_DATE", "2", "KOGAS_W_NAME_DATE", "Date", "Date"),
            ("Dwg No.:", "KOGAS_W_COL_DWG", "3", "KOGAS_W_NAME_DWG", "Drawing No.", "Dwg"),
            ("Joint No.:", "KOGAS_W_COL_JOINT", "4", "KOGAS_W_NAME_JOINT", "Film Ident. No.", "Joint"),
            ("Location:", "KOGAS_W_COL_LOC", "5", "KOGAS_W_NAME_LOC", "Film Location", "Loc"),
            ("T:", "KOGAS_W_COL_THK", "6", "KOGAS_W_NAME_THK", "T", "T"),
            ("Mat:", "KOGAS_W_COL_MAT", "7", "KOGAS_W_NAME_MAT", "Mat", "Mat"),
            ("판정(Result):", "KOGAS_W_COL_RES", "28", "KOGAS_W_NAME_RES", "Result", "Result"),
            ("용접사:", "KOGAS_W_COL_WELDER", "29", "KOGAS_W_NAME_WELDER", "Welder No", "Welder"),
            ("비고:", "KOGAS_W_COL_REM", "30", "KOGAS_W_NAME_REM", "Remarks", "Remarks"),
            ("① Crack:", "KOGAS_W_COL_D1", "17", "KOGAS_W_NAME_D1", "① Crack", "D1"),
            ("② IP:", "KOGAS_W_COL_D2", "18", "KOGAS_W_NAME_D2", "② IP", "D2"),
            ("③ LF:", "KOGAS_W_COL_D3", "19", "KOGAS_W_NAME_D3", "③ LF", "D3"),
            ("④ Slag:", "KOGAS_W_COL_D4", "20", "KOGAS_W_NAME_D4", "④ Slag", "D4"),
            ("⑤ Por:", "KOGAS_W_COL_D5", "21", "KOGAS_W_NAME_D5", "⑤ Por", "D5"),
            ("⑥ U/C:", "KOGAS_W_COL_D6", "22", "KOGAS_W_NAME_D6", "⑥ U/C", "D6"),
            ("⑦ RUC:", "KOGAS_W_COL_D7", "23", "KOGAS_W_NAME_D7", "⑦ RUC", "D7"),
            ("⑧ BT:", "KOGAS_W_COL_D8", "24", "KOGAS_W_NAME_D8", "⑧ BT", "D8"),
            ("⑨ IC:", "KOGAS_W_COL_D9", "25", "KOGAS_W_NAME_D9", "⑨ IC", "D9"),
            ("⑩ CP:", "KOGAS_W_COL_D10", "26", "KOGAS_W_NAME_D10", "⑩ CP", "D10"),
            ("⑪ RC:", "KOGAS_W_COL_D11", "27", "KOGAS_W_NAME_D11", "⑪ RC", "D11"),
            ("⑫ Mis:", "KOGAS_W_COL_D12", "28", "KOGAS_W_NAME_D12", "⑫ Mis", "D12"),
            ("⑬ EP:", "KOGAS_W_COL_D13", "29", "KOGAS_W_NAME_D13", "⑬ EP", "D13"),
            ("⑭ SD:", "KOGAS_W_COL_D14", "30", "KOGAS_W_NAME_D14", "⑭ SD", "D14"),
            ("⑮ Oth:", "KOGAS_W_COL_D15", "31", "KOGAS_W_NAME_D15", "⑮ Oth", "D15")
        ]

        render_fields(read_frame, read_items, is_read=True)
        render_fields(write_frame, write_items, is_read=False)

        ttk.Label(scrollable_frame, text="* 엑셀열: 알파벳(A, B...) 또는 숫자(1, 2...) 입력 (0=제외)", foreground="gray", font=("Malgun Gothic", 8)).pack(pady=10)

    def _update_mode_heading(self, mode, internal_id, var, key_name):
        """[NEW] 특정 모드의 미리보기 헤더 이름을 실시간으로 업데이트 및 설정 저장"""
        try:
            new_text = var.get()
            # 1. Update config immediately so it survives refreshes
            self.config[key_name] = new_text
            
            # 2. Update existing tree UI
            tree = None
            if mode == "PMI": tree = self.preview_tree
            elif mode == "RT": tree = self.rt_preview_tree
            elif mode == "KOGAS": tree = self.kogas_preview_tree
            elif mode == "PT": tree = self.pt_preview_tree
            
            if tree:
                try: tree.heading(internal_id, text=new_text)
                except: pass
        except Exception:
            pass

    def _create_row_settings(self, parent, mode=None):
        if mode is None or mode == "PMI":
            # PMI Rows
            pmi_frame = ttk.LabelFrame(parent, text=" PMI 행 설정 ", padding=10)
            pmi_frame.pack(fill='x', pady=5)
            pmi_rows = [
                ("을지 데이터 시작", "START_ROW", "PMI 데이터 시작"), 
                ("을지 데이터 종료", "DATA_END_ROW", "PMI 데이터 종료"), 
                ("을지 인쇄 영역 끝", "PRINT_END_ROW", "PMI 인쇄 끝"),
                ("갑지 데이터 시작", "GAPJI_START_ROW", "PMI 갑지 데이터 시작/테두리"),
                ("갑지 데이터 종료", "GAPJI_DATA_END_ROW", "PMI 갑지 데이터 종료/테두리"),
                ("갑지 인쇄 영역 끝", "GAPJI_PRINT_END_ROW", "PMI 갑지 인쇄 끝")
            ]
            for i, (label, key, tip) in enumerate(pmi_rows):
                ttk.Label(pmi_frame, text=label + ":").grid(row=i, column=0, sticky='e', padx=5, pady=2)
                # Default values for Gapji settings
                def_val = "0"
                if key == "GAPJI_PRINT_END_ROW": def_val = "51"
                elif key == "GAPJI_START_ROW": def_val = "23"
                elif key == "GAPJI_DATA_END_ROW": def_val = "38"
                else: def_val = str(self.config.get(key, 0))
                
                v = tk.StringVar(value=str(self.config.get(key, def_val)))
                ent = ttk.Entry(pmi_frame, textvariable=v, width=10)
                ent.grid(row=i, column=1, sticky='w', padx=5)
                ent.bind("<Return>", lambda e: self.root.focus_set())
                self.setting_vars[key] = v
                ttk.Label(pmi_frame, text=f"({tip})", foreground="gray", font=("Malgun Gothic", 9)).grid(row=i, column=2, sticky='w')

        if mode is None or mode == "PAUT":
            # PAUT Rows
            ut_row_frame = ttk.LabelFrame(parent, text=" PAUT 행 설정 ", padding=10)
            ut_row_frame.pack(fill='x', pady=5)
            ut_rows = [
                ("을지 데이터 시작", "PAUT_START_ROW", "PAUT 데이터 시작"), 
                ("을지 데이터 종료", "PAUT_DATA_END_ROW", "PAUT 데이터 종료"), 
                ("을지 인쇄 영역 끝", "PAUT_PRINT_END_ROW", "PAUT 인쇄 끝"),
                ("갑지 데이터 시작", "PAUT_GAPJI_START_ROW", "PAUT 갑지 데이터 시작"),
                ("갑지 데이터 종료", "PAUT_GAPJI_DATA_END_ROW", "PAUT 갑지 데이터 종료"),
                ("갑지 인쇄 영역 끝", "PAUT_GAPJI_PRINT_END_ROW", "PAUT 갑지 끝")
            ]
            for i, (label, key, tip) in enumerate(ut_rows):
                ttk.Label(ut_row_frame, text=label + ":").grid(row=i, column=0, sticky='e', padx=5, pady=2)
                def_val = "51" if "GAPJI_PRINT" in key else "0"
                v = tk.StringVar(value=str(self.config.get(key, def_val)))
                ent = ttk.Entry(ut_row_frame, textvariable=v, width=10)
                ent.grid(row=i, column=1, sticky='w', padx=5)
                ent.bind("<Return>", lambda e: self.root.focus_set())
                self.setting_vars[key] = v
                ttk.Label(ut_row_frame, text=f"({tip})", foreground="gray", font=("Malgun Gothic", 9)).grid(row=i, column=2, sticky='w')

        if mode is None or mode in ["RT", "KOGAS"]:
            # RT/KOGAS Rows
            m_label = "KOGAS" if mode == "KOGAS" else "RT"
            rt_row_frame = ttk.LabelFrame(parent, text=f" {m_label} 행 설정 ", padding=10)
            rt_row_frame.pack(fill='x', pady=5)
            
            # Use mode-specific keys
            m_key = "KOGAS" if mode == "KOGAS" else "RT"
            rt_rows = [
                ("을지 데이터 시작", f"{m_key}_START_ROW", f"{m_label} 데이터 시작"), 
                ("을지 데이터 종료", f"{m_key}_DATA_END_ROW", f"{m_label} 데이터 종료"), 
                ("을지 인쇄 영역 끝", f"{m_key}_PRINT_END_ROW", f"{m_label} 인쇄 끝"),
                ("갑지 데이터 시작", f"{m_key}_GAPJI_START_ROW", f"{m_label} 갑지 데이터 시작"),
                ("갑지 데이터 종료", f"{m_key}_GAPJI_DATA_END_ROW", f"{m_label} 갑지 데이터 종료"),
                ("갑지 인쇄 영역 끝", f"{m_key}_GAPJI_PRINT_END_ROW", f"{m_label} 갑지 끝")
            ]
            for i, (label, key, tip) in enumerate(rt_rows):
                ttk.Label(rt_row_frame, text=label + ":").grid(row=i, column=0, sticky='e', padx=5, pady=2)
                def_val = "51" if "GAPJI_PRINT" in key else "0"
                v = tk.StringVar(value=str(self.config.get(key, def_val)))
                ent = ttk.Entry(rt_row_frame, textvariable=v, width=10)
                ent.grid(row=i, column=1, sticky='w', padx=5)
                ent.bind("<Return>", lambda e: self.root.focus_set())
                self.setting_vars[key] = v
                ttk.Label(rt_row_frame, text=f"({tip})", foreground="gray", font=("Malgun Gothic", 9)).grid(row=i, column=2, sticky='w')

        if mode is None or mode == "PT":
            # PT Rows
            pt_row_frame = ttk.LabelFrame(parent, text=" PT 행 설정 ", padding=10)
            pt_row_frame.pack(fill='x', pady=5)
            pt_rows = [
                ("을지 데이터 시작", "PT_START_ROW", "PT 데이터 시작"), 
                ("을지 데이터 종료", "PT_DATA_END_ROW", "PT 데이터 종료"), 
                ("을지 인쇄 영역 끝", "PT_PRINT_END_ROW", "PT 인쇄 끝"),
                ("갑지 데이터 시작", "PT_GAPJI_START_ROW", "PT 갑지 데이터 시작"),
                ("갑지 데이터 종료", "PT_GAPJI_DATA_END_ROW", "PT 갑지 데이터 종료"),
                ("갑지 인쇄 영역 끝", "PT_GAPJI_PRINT_END_ROW", "PT 갑지 끝")
            ]
            for i, (label, key, tip) in enumerate(pt_rows):
                ttk.Label(pt_row_frame, text=label + ":").grid(row=i, column=0, sticky='e', padx=5, pady=2)
                # Default values for PT
                if key == "PT_START_ROW": def_val = "18"
                elif key == "PT_DATA_END_ROW": def_val = "32"
                elif key == "PT_PRINT_END_ROW": def_val = "35"
                elif key == "PT_GAPJI_PRINT_END_ROW": def_val = "35"
                elif "GAPJI_START" in key: def_val = "0"
                elif "GAPJI_DATA_END" in key: def_val = "35"
                else: def_val = "0"
                
                v = tk.StringVar(value=str(self.config.get(key, def_val)))
                ent = ttk.Entry(pt_row_frame, textvariable=v, width=10)
                ent.grid(row=i, column=1, sticky='w', padx=5)
                ent.bind("<Return>", lambda e: self.root.focus_set())
                self.setting_vars[key] = v
                ttk.Label(pt_row_frame, text=f"({tip})", foreground="gray", font=("Malgun Gothic", 9)).grid(row=i, column=2, sticky='w')

    def _create_preview_ui(self, parent):
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # [NEW] File Info Header (Contextual Awareness)
        header_info = tk.Frame(container, background="#ffffff", highlightthickness=1, highlightbackground="#e5e7eb")
        header_info.pack(fill='x', pady=(0, 5))
        tk.Label(header_info, textvariable=self.file_info_vars['PMI'], background="#ffffff", 
                 foreground="#4b5563", font=("Malgun Gothic", 8, "bold"), padx=10, pady=2).pack(side='left')

        # [REFINED] Compact Header inside Frame
        header_frame = tk.Frame(container, background="#f9fafb")
        header_frame.pack(fill='x', pady=(0, 5))
        
        tk.Checkbutton(header_frame, text="⚠️ 함량 미달만 보기(PMI)", variable=self.pmi_show_deficiency_only, 
                       background="#f9fafb", font=("Malgun Gothic", 9),
                       command=lambda: self.populate_preview(self.extracted_data, switch_tab=False)).pack(side='left')

        tree_frame = tk.Frame(container, background="#f9fafb")
        self.preview_tree = ttk.Treeview(tree_frame, columns=("ST", "V", "No", "Date", "Drawing No.", "Joint No.", "Location", "Ni", "Cr", "Mo", "Result"), show='headings', height=10, selectmode='extended')
        # [NEW] Highlight tags
        self.preview_tree.tag_configure("deficient", background="#fee2e2", foreground="#991b1b") # Light red
        self.preview_tree.tag_configure("group_even", background="#ffffff")
        self.preview_tree.tag_configure("group_odd", background="#f3f4f6")
        
        saved_widths = self.config.get("PMI_COL_WIDTHS", {})
        default_widths = {"ST": 40, "V": 40, "No": 50, "Date": 90, "Drawing No.": 400, "Joint No.": 200, "Location": 300, "Ni": 60, "Cr": 60, "Mo": 60, "Result": 150}
        
        for col in self.preview_tree["columns"]:
            # Use dynamic names from config for headings
            name_key = f"PMI_NAME_{col.split('(')[0].replace(' ', '').replace('.', '').replace('-', '').upper()}"
            if col == "No": name_key = "PMI_NAME_NO"
            elif col == "Date": name_key = "PMI_NAME_DATE"
            elif col == "Drawing No.": name_key = "PMI_NAME_DWG"
            elif col == "Joint No.": name_key = "PMI_NAME_JOINT"
            elif col == "Location": name_key = "PMI_NAME_LOC"
            elif col == "Ni": name_key = "PMI_NAME_NI"
            elif col == "Cr": name_key = "PMI_NAME_CR"
            elif col == "Mo": name_key = "PMI_NAME_MO"
            elif col == "Result": name_key = "PMI_NAME_RES"
            else: name_key = None
            
            display_text = self.config.get(name_key, col) if name_key else col
            self.preview_tree.heading(col, text=display_text, anchor='center', command=lambda _c=col: self.sort_by_column(_c, mode="PMI"))
            w = saved_widths.get(col, default_widths.get(col, 80))
            self.preview_tree.column(col, width=w, anchor='center', stretch=False)
        
        # [NEW] Add Scrollbars for full coverage
        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.preview_tree.yview)
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.preview_tree.grid(row=0, column=0, sticky='nsew')
        scroll_y.grid(row=0, column=1, sticky='ns')
        scroll_x.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        self._setup_preview_sidebar(self.preview_tree, container, mode="PMI")
        tree_frame.pack(side="left", fill="both", expand=True)

    def _create_rt_preview_ui(self, parent, mode="RT"):
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # [NEW] File Info Header
        header_info = tk.Frame(container, background="#ffffff", highlightthickness=1, highlightbackground="#e5e7eb")
        header_info.pack(fill='x', pady=(0, 5))
        tk.Label(header_info, textvariable=self.file_info_vars.get(mode, tk.StringVar()), background="#ffffff", 
                 foreground="#4b5563", font=("Malgun Gothic", 8, "bold"), padx=10, pady=2).pack(side='left')

        # Inner frame for horizontal/vertical scroll
        tree_frame = tk.Frame(container, background="#f9fafb")
        tree = ttk.Treeview(tree_frame, columns=[], show='headings', height=10, selectmode='extended')
        if mode == "RT": self.rt_preview_tree = tree
        else: self.kogas_preview_tree = tree

        # [DYNAMIC] 가스공사 모드 여부에 따른 초기 컬럼 설정
        self._update_rt_preview_columns(mode=mode)
        
        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        tree.grid(row=0, column=0, sticky='nsew')
        scroll_y.grid(row=0, column=1, sticky='ns')
        scroll_x.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        self._setup_preview_sidebar(tree, container, mode=mode)
        tree_frame.pack(side="left", fill="both", expand=True)

    def _setup_preview_sidebar(self, tree, container, mode):
        # 1. Event Bindings
        self.drag_start_item = None
        
        def _on_tree_press(event, t=tree, m=mode):
            t.focus_set()
            
            # Column Drag-and-Drop Init
            if t.identify_region(event.x, event.y) == "heading":
                # Ensure we are not grabbing the separator
                if t.identify_region(event.x - 3, event.y) == "separator" or t.identify_region(event.x + 3, event.y) == "separator":
                    t.drag_col_id = None
                else:
                    t.drag_col_id = t.identify_column(event.x)
                return
            else:
                t.drag_col_id = None

            item_id = t.identify_row(event.y)
            col_id = t.identify_column(event.x)
            
            # [DYNAMIC] Get consistent module info
            mode_info = self._get_mode_info(m)
            if not mode_info: return
            _, idx_map, data_list, col_keys = mode_info
            
            self.start_col = col_id
            self.end_col = col_id
            self.last_clicked_col = col_id
            
            if not item_id or not col_id: return
            
            try:
                col_idx = int(col_id.replace("#", "")) - 1
                if 0 <= col_idx < len(col_keys):
                    key = col_keys[col_idx]
                    # Checkbox Toggle (V or selected)
                    if key in ["selected", "V"] or col_id == "#1": # Fallback to first col for V
                        if key == "selected" or (m == "PMI" and col_id == "#2"):
                             view_idx = t.index(item_id)
                             if 0 <= view_idx < len(idx_map):
                                 actual_idx = idx_map[view_idx]
                                 data_list[actual_idx]['selected'] = not data_list[actual_idx].get('selected', True)
                                 self.populate_preview(data_list, switch_tab=False, mode=m)
                                 return "break"
                    
                    # RT Specific specialized toggle
                    if m == "RT" and ((key.startswith("D") and key[1:].isdigit()) or (key in ["Acc", "Rej"])):
                        view_idx = t.index(item_id)
                        if 0 <= view_idx < len(idx_map):
                            actual_idx = idx_map[view_idx]
                            old_v = data_list[actual_idx].get(key, "")
                            new_v = "√" if old_v == "" else ""
                            data_list[actual_idx][key] = new_v
                            if key in ["Acc", "Rej"] and new_v == "√":
                                other = "Rej" if key == "Acc" else "Acc"
                                data_list[actual_idx][other] = ""
                            self.populate_preview(data_list, switch_tab=False, mode="RT")
                            return "break"
            except: pass

            self.drag_start_item = item_id
            if not (event.state & 0x0001 or event.state & 0x0004):
                t.selection_set(item_id)

        def _on_tree_drag(event, t=tree):
            if getattr(t, 'drag_col_id', None):
                return
                
            if not self.drag_start_item: return
            curr_item = t.identify_row(event.y)
            drag_col = t.identify_column(event.x)
            if drag_col: self.end_col = drag_col
            
            if not curr_item: return
            all_items = t.get_children('')
            try:
                low = min(all_items.index(self.drag_start_item), all_items.index(curr_item))
                high = max(all_items.index(self.drag_start_item), all_items.index(curr_item))
                t.selection_set(all_items[low:high+1])
            except: pass

        def _on_tree_release(event, t=tree):
            self.drag_start_item = None
            
            if getattr(t, 'drag_col_id', None):
                target_col = t.identify_column(event.x)
                if target_col and target_col != t.drag_col_id:
                    try:
                        curr_display = list(t["displaycolumns"])
                        if curr_display == ["#all"]:
                            curr_display = list(t["columns"])
                            
                        drag_id = t.column(t.drag_col_id, "id")
                        target_id = t.column(target_col, "id")
                        
                        if drag_id in curr_display and target_id in curr_display:
                            curr_display.remove(drag_id)
                            target_idx = curr_display.index(target_id)
                            curr_display.insert(target_idx, drag_id)
                            t["displaycolumns"] = tuple(curr_display)
                    except Exception:
                        pass
                t.drag_col_id = None

        tree.bind("<Button-1>", _on_tree_press)
        tree.bind("<B1-Motion>", _on_tree_drag)
        tree.bind("<ButtonRelease-1>", _on_tree_release)
        tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        tree.bind("<Double-1>", lambda e, m=mode: self.on_tree_double_click(e, m))
        tree.bind("<Button-3>", lambda e, m=mode: self.show_context_menu(e, m))
        tree.bind("<Control-c>", lambda e, m=mode: self.copy_cell(m))
        tree.bind("<Control-C>", lambda e, m=mode: self.copy_cell(m))
        tree.bind("<Control-v>", lambda e, m=mode: self.paste_cell(m))
        tree.bind("<Control-V>", lambda e, m=mode: self.paste_cell(m))
        tree.bind("<Control-a>", lambda e, t=tree: [t.selection_set(t.get_children(''))])
        tree.bind("<Control-A>", lambda e, t=tree: [t.selection_set(t.get_children(''))])
        tree.bind("<Delete>", lambda e, m=mode: self.on_delete_key(e, m))
        tree.bind("<F2>", lambda e, m=mode: self.on_tree_double_click(e, m))

        # 2. Sidebar with SCROLL support (too many elements to fit without scrolling)
        sidebar_outer = tk.Frame(container, background="#f1f5f9", highlightthickness=1, highlightbackground="#e2e8f0", width=160)
        sidebar_outer.pack(side='left', fill='y', padx=(0, 5))
        sidebar_outer.pack_propagate(False)
        
        # --- STICKY TOP FILTERS ---
        sticky_top_frame = tk.Frame(sidebar_outer, background="#f1f5f9", padx=5, pady=3)
        sticky_top_frame.pack(side='top', fill='x')

        ttk.Label(sticky_top_frame, text="날짜 필터", font=("Malgun Gothic", 8, "bold")).pack(pady=(2, 1))
        date_frame = tk.Frame(sticky_top_frame, background="#f1f5f9")
        date_frame.pack(fill='x', pady=2)
        
        listbox = tk.Listbox(date_frame, selectmode='single', height=5, exportselection=False, font=("Malgun Gothic", 8))
        listbox.pack(side='left', fill='both', expand=True)
        sb = ttk.Scrollbar(date_frame, orient='vertical', command=listbox.yview)
        sb.pack(side='right', fill='y')
        listbox.config(yscrollcommand=sb.set)
        
        if mode == "RT": self.rt_date_listbox = listbox
        elif mode == "PT": self.pt_date_listbox = listbox
        elif mode == "PAUT": self.paut_date_listbox = listbox
        elif mode == "KOGAS": self.kogas_date_listbox = listbox
        else: self.date_listbox = listbox
        
        def _toggle_date(event, lb=listbox):
            idx = lb.nearest(event.y)
            if idx >= 0:
                val = lb.get(idx)
                if val.startswith("[v]"):
                    lb.delete(idx); lb.insert(idx, val.replace("[v] ", "[ ] "))
                elif val.startswith("[ ]"):
                    lb.delete(idx); lb.insert(idx, val.replace("[ ] ", "[v] "))
                lb.selection_clear(0, tk.END)
        
        listbox.bind("<ButtonRelease-1>", _toggle_date)
        
        def _apply_date_filter(m=mode, lb=listbox):
            sel_dates = [lb.get(i).replace("[v] ", "").replace("[ ] ", "") for i in range(lb.size()) if lb.get(i).startswith("[v]")]
            if m == "RT": data = self.rt_extracted_data
            elif m == "KOGAS": data = self.kogas_extracted_data
            elif m == "PT": data = self.pt_extracted_data
            elif m == "PAUT": data = self.paut_extracted_data
            else: data = self.extracted_data
            
            for item in data:
                item['date_filtered'] = (item.get('Date', '') in sel_dates)
            
            if m == "PMI":
                self.update_pmi_loc_listbox()
                
            self.populate_preview(data, switch_tab=False, mode=m)
            
        ttk.Button(sticky_top_frame, text="날짜 적용", command=_apply_date_filter).pack(fill='x', pady=(0, 5))
        
        # [NEW] Multi-Select Test Location Filter (PMI Only)
        if mode == "PMI":
            ttk.Label(sticky_top_frame, text="지점(Loc) 필터", font=("Malgun Gothic", 8, "bold"), background="#f1f5f9").pack(pady=(2, 1))
            loc_frame = tk.Frame(sticky_top_frame, background="#f1f5f9")
            loc_frame.pack(fill='x', pady=2)
            
            self.pmi_loc_listbox = tk.Listbox(loc_frame, selectmode='single', height=4, exportselection=False, font=("Malgun Gothic", 8))
            self.pmi_loc_listbox.pack(side='left', fill='both', expand=True)
            sb_loc = ttk.Scrollbar(loc_frame, orient='vertical', command=self.pmi_loc_listbox.yview)
            sb_loc.pack(side='right', fill='y')
            self.pmi_loc_listbox.config(yscrollcommand=sb_loc.set)
            
            def _toggle_loc(event):
                idx = self.pmi_loc_listbox.nearest(event.y)
                if idx >= 0:
                    val = self.pmi_loc_listbox.get(idx)
                    if val.startswith("[v]"):
                        self.pmi_loc_listbox.delete(idx); self.pmi_loc_listbox.insert(idx, val.replace("[v] ", "[ ] "))
                    elif val.startswith("[ ]"):
                        self.pmi_loc_listbox.delete(idx); self.pmi_loc_listbox.insert(idx, val.replace("[ ] ", "[v] "))
                    self.pmi_loc_listbox.selection_clear(0, tk.END)
            
            self.pmi_loc_listbox.bind("<ButtonRelease-1>", _toggle_loc)
            
            def _apply_loc_filter():
                self.populate_preview(self.extracted_data, switch_tab=False, mode="PMI")
            
            ttk.Button(sticky_top_frame, text="지점 적용", command=_apply_loc_filter).pack(fill='x', pady=(0, 5))

        # [NEW] Sticky 'Show Selected Only' Checkbox
        tk.Checkbutton(sticky_top_frame, text="선택만", variable=self.show_selected_only, 
                       command=lambda: self.populate_preview(
                           self.rt_extracted_data if mode == "RT" else (
                               self.kogas_extracted_data if mode == "KOGAS" else (
                                   self.pt_extracted_data if mode == "PT" else (
                                       self.paut_extracted_data if mode == "PAUT" else self.extracted_data
                                   )
                               )
                           ), switch_tab=False, mode=mode),
                       background="#f1f5f9", font=("Malgun Gothic", 8)).pack(pady=(0, 3), anchor='w')

        # --- SCROLLABLE BUTTONS AREA ---
        scroll_wrapper = tk.Frame(sidebar_outer, background="#f1f5f9")
        scroll_wrapper.pack(side='top', fill='both', expand=True)

        sidebar_canvas = tk.Canvas(scroll_wrapper, background="#f1f5f9", highlightthickness=0)
        sidebar_sb = ttk.Scrollbar(scroll_wrapper, orient="vertical", command=sidebar_canvas.yview)
        sidebar = tk.Frame(sidebar_canvas, background="#f1f5f9", padx=5, pady=3)
        
        sidebar.bind("<Configure>", lambda e: sidebar_canvas.configure(scrollregion=sidebar_canvas.bbox("all")))
        # Fixed width for inner sidebar relative to outer width (160)
        # Scrollbar takes about ~15px, let's make window ~140px
        sidebar_canvas.create_window((0, 0), window=sidebar, anchor="nw", width=140)
        sidebar_canvas.configure(yscrollcommand=sidebar_sb.set)
        
        sidebar_canvas.pack(side='left', fill='both', expand=True)
        sidebar_sb.pack(side='right', fill='y')
        
        # Mousewheel scroll for sidebar
        def _sidebar_scroll(event):
            # Bulletproof bounds checking to ensure no overscroll detach
            top, bottom = sidebar_canvas.yview()
            direction = int(-1*(event.delta/120))
            
            # If everything fits, lock it
            if top <= 0.0 and bottom >= 1.0:
                sidebar_canvas.yview_moveto(0)
                return "break"
                
            if direction < 0 and top <= 0.0:
                sidebar_canvas.yview_moveto(0)
                return "break"
                
            if direction > 0 and bottom >= 1.0:
                sidebar_canvas.yview_moveto(1.0)
                return "break"
                
            # Perform scroll
            sidebar_canvas.yview_scroll(direction, "units")
            
            # Post-scroll enforcement for overshoot
            new_top, new_bottom = sidebar_canvas.yview()
            if direction < 0 and new_top <= 0.0:
                sidebar_canvas.yview_moveto(0)
            elif direction > 0 and new_bottom >= 1.0:
                sidebar_canvas.yview_moveto(1.0)
                
            return "break"
            
        sidebar_canvas.bind("<MouseWheel>", _sidebar_scroll)
        sidebar.bind("<MouseWheel>", _sidebar_scroll)
        def _bind_mousewheel_recursive(widget):
            widget.bind("<MouseWheel>", _sidebar_scroll)
            for child in widget.winfo_children():
                _bind_mousewheel_recursive(child)
        sidebar.bind("<Configure>", lambda e: _bind_mousewheel_recursive(sidebar), add="+")

        ttk.Button(sidebar, text="전체 선택", width=12, command=lambda: self.select_all(mode)).pack(pady=2, fill='x')
        ttk.Button(sidebar, text="선택 해제", width=12, command=lambda: self.deselect_all(mode)).pack(pady=2, fill='x')
        
        if mode == "PMI":
            ttk.Button(sidebar, text="미달만 선택", width=12, command=lambda: self.select_deficient_items()).pack(pady=2, fill='x')
        
        ttk.Button(sidebar, text="컬럼 관리", width=12, command=lambda: self.manage_columns(mode)).pack(pady=2, fill='x')
        
        if mode == "PAUT":
            ttk.Button(sidebar, text="행 추가", width=12, command=lambda: self.add_item(mode)).pack(pady=2, fill='x')
        
        ttk.Button(sidebar, text="병합", width=12, command=lambda: self.merge_selected_iso(mode)).pack(pady=2, fill='x')
        ttk.Button(sidebar, text="일괄 변경", width=12, command=lambda: self.show_bulk_update_dialog(mode)).pack(pady=(2, 3), fill='x')

        tk.Frame(sidebar, height=1, background="#e5e7eb").pack(fill='x', pady=5)
        ttk.Button(sidebar, text="▲ 위로", width=10, command=lambda: self.move_item(-1, mode)).pack(pady=5)
        ttk.Button(sidebar, text="▼ 아래로", width=10, command=lambda: self.move_item(1, mode)).pack(pady=5)
        tk.Frame(sidebar, height=10, background="#f9fafb").pack()
        ttk.Button(sidebar, text="선택 삭제", width=12, command=lambda: self.delete_item(mode)).pack(pady=5)
        ttk.Button(sidebar, text="전체 초기화", width=12, command=lambda: self.clear_all(mode)).pack(pady=(2, 10))

        tk.Frame(sidebar, height=1, background="#e5e7eb").pack(fill='x', pady=5)
        ttk.Button(sidebar, text="💾 현재 내용 저장", width=18, command=lambda: self.save_preview_data(mode)).pack(pady=5)
        ttk.Button(sidebar, text="📂 저장된 내용 열기", width=18, command=lambda: self.load_preview_data(mode)).pack(pady=5)
        ttk.Button(sidebar, text="📊 엑셀 파일로 추출", width=18, command=lambda: self.export_to_excel(mode)).pack(pady=(10, 2))

        # 3. Context Menu Config
        for t in ["evenrow", "oddrow", "group_even", "group_odd", "grouped_even", "grouped_odd", "item_even", "item_odd"]:
            tree.tag_configure(t, background="", foreground="black")
        tree.tag_configure('group_even', font=("Malgun Gothic", 10, "bold"))
        tree.tag_configure('group_odd', font=("Malgun Gothic", 10, "bold"))


    def on_tree_select(self, event):
        """항목 선택 시 체크박스 상태와 동기화 (현재는 드래그 안정성을 위해 드래그 중에는 무시)"""
        pass

    def toggle_item_selection(self, item_id, mode="PMI"):
        """단일 항목의 체크 상태를 토글"""
        tree = self.rt_preview_tree if mode == "RT" else self.preview_tree
        idx_map = self.rt_item_idx_map if mode == "RT" else self.item_idx_map
        data = self.rt_extracted_data if mode == "RT" else self.extracted_data
        
        view_idx = tree.index(item_id)
        if 0 <= view_idx < len(idx_map):
            actual_idx = idx_map[view_idx]
            data[actual_idx]['selected'] = not data[actual_idx].get('selected', True)
            self.populate_preview(data, switch_tab=False, mode=mode)
            
            # 선택 상태 복구
            num_items = len(tree.get_children())
            if view_idx < num_items:
                new_item = tree.get_children()[view_idx]
                tree.selection_set(new_item)
                tree.focus(new_item)

    def toggle_selected_items(self, mode="PMI"):
        """선택된 모든 항목의 체크 상태를 토글"""
        tree = self.rt_preview_tree if mode == "RT" else self.preview_tree
        idx_map = self.rt_item_idx_map if mode == "RT" else self.item_idx_map
        data = self.rt_extracted_data if mode == "RT" else self.extracted_data
        
        selected_ids = tree.selection()
        if not selected_ids: return
        
        for item_id in selected_ids:
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                data[actual_idx]['selected'] = not data[actual_idx].get('selected', True)
        
        self.populate_preview(data, switch_tab=False, mode=mode)
        # 선택 보존
        for sid in selected_ids:
            try: tree.selection_add(sid)
            except: pass

    def on_tree_click_checkbox(self, event):
        """Reserved for future coordinate-based checkbox toggle if needed without blocking drag."""
        pass

    def on_tree_double_click(self, event, mode="PMI"):
        """셀(Cell) 더블클릭 시 엑셀처럼 그 자리에서 바로 수동 편집 가능하게 함"""
        mode_info = self._get_mode_info(mode)
        if not mode_info: return
        tree, idx_map, data_list, col_keys = mode_info

        region = tree.identify_region(event.x, event.y)
        if region != "cell": return
        
        column = tree.identify_column(event.x)
        item_id = tree.identify_row(event.y)
        if not item_id or not column: return
        
        # [DYNAMIC] 컬럼 인덱스를 통해 키 추출
        try:
            col_idx = int(column.replace('#', '')) - 1
            if col_idx < 0 or col_idx >= len(col_keys): return
            key = col_keys[col_idx]
        except: return
        
        # 수정 차단 컬럼 (V, 상태 아이콘 등)
        if key in ["selected", "_status", "V", "ST"]:
            if key in ["selected", "V"]: self.toggle_item_selection(item_id, mode)
            return

        # 현재 값 가져오기
        x, y, w, h = tree.bbox(item_id, column)
        view_idx = tree.index(item_id)
        if not (0 <= view_idx < len(idx_map)): return
        actual_idx = idx_map[view_idx]
        
        curr_val = data_list[actual_idx].get(key, "")
        numeric_keys = ["Th'k(mm)", "Start", "End", "Length(mm)", "Upper", "Lower", "Height(mm)", "a/l", "a/t"]
        if (key in numeric_keys or (key not in ["No", "Joint No.", "Line No.", "Joint", "Loc", "Grade", "Date", "Dwg", "ISO", "ISO/DWG", "Type of Flaw", "Evaluation", "Remarks", "Result", "Location", "Nature", "selected", "_status", "ST", "V"])) and self.to_float(curr_val) > 0:
            curr_val = f"{self.to_float(curr_val):.2f}"
            
        entry = ttk.Entry(tree, font=("Malgun Gothic", 10), exportselection=True)
        entry.insert(0, str(curr_val))
        entry.select_range(0, tk.END)
        entry.place(x=x, y=y, width=w, height=h)
        entry.focus_set()
        
        def finish_edit(e=None):
            if not entry.winfo_exists(): return
            new_val = entry.get().strip()
            
            # [PMI 특화] 함량 소수점 처리 및 판정 재계산
            if mode == "PMI" and key not in ["No", "Joint", "Loc", "Grade", "Date", "Dwg", "_status", "ST", "V", "selected", "order_index"]:
                f_val = self.to_float(new_val)
                data_list[actual_idx][key] = f"{f_val:.2f}" if f_val > 0 else ""
                # 함량 변경 시 등급 판정 재수행
                new_grade = self.check_material_grade(data_list[actual_idx])
                if new_grade: data_list[actual_idx]['Grade'] = new_grade
            elif mode == "PAUT" and key in ["Th'k(mm)", "Height(mm)", "Length(mm)", "Type of Flaw"]:
                data_list[actual_idx][key] = new_val
                # [PAUT] Recalculate Evaluation and Ratios
                t = self.to_float(data_list[actual_idx].get('Th\'k(mm)', 0))
                h = self.to_float(data_list[actual_idx].get('Height(mm)', 0))
                l = self.to_float(data_list[actual_idx].get('Length(mm)', 0))
                d = self.to_float(data_list[actual_idx].get('Upper', 0))
                nat = data_list[actual_idx].get('Type of Flaw', 'Slag')
                eval_mode = self.paut_eval_mode.get()
                
                res, _ = self.evaluate_paut_flaw(t, h, l, d, nat, mode=eval_mode)
                data_list[actual_idx]['Evaluation'] = res
                if l > 0: data_list[actual_idx]['a/l'] = f"{h/l:.3f}"
                if t > 0: data_list[actual_idx]['a/t'] = f"{h/t:.3f}"
            else:
                data_list[actual_idx][key] = new_val

            self.populate_preview(data_list, switch_tab=False, mode=mode)
            self.log(f"📝 {mode} {key} 수정: {new_val}")
            entry.destroy()

        entry.bind("<Return>", finish_edit)
        entry.bind("<FocusOut>", lambda e: finish_edit())
        entry.bind("<Escape>", lambda e: entry.destroy())

    def on_delete_key(self, event, mode="PMI"):
        """Delete 키 선택 시 현재 선택된 셀(들)의 내용을 지움 (엑셀 방식)"""
        mode_info = self._get_mode_info(mode)
        if not mode_info: return
        tree, idx_map, data_list, col_keys = mode_info
        
        selected = tree.selection()
        if not selected: return
        
        # 드래그된 컬럼 범위 계산
        s_col = getattr(self, 'start_col', None)
        e_col = getattr(self, 'end_col', None)
        l_col = getattr(self, 'last_clicked_col', None)
        
        if not s_col or not e_col or s_col == e_col:
            target_col_ids = [l_col or s_col or "#4"]
        else:
            try:
                s_idx = int(s_col.replace('#', '')) - 1
                e_idx = int(e_col.replace('#', '')) - 1
                col_start, col_end = min(s_idx, e_idx), max(s_idx, e_idx)
                target_col_ids = [f"#{i+1}" for i in range(col_start, col_end + 1)]
            except:
                target_col_ids = [l_col or s_col or "#4"]
        
        key_map = {f"#{i+1}": k for i, k in enumerate(col_keys)}
        count = 0
        for item_id in selected:
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                for c_id in target_col_ids:
                    key = key_map.get(c_id)
                    if key and key not in ["selected", "_status"]:
                        data_list[actual_idx][key] = ""
                        count += 1
        
        if count > 0:
            self.populate_preview(data_list, switch_tab=False, mode=mode)
            self.log(f"🧹 {mode} 데이터 {count}건 내용 삭제 완료")

    def show_context_menu(self, event, mode="PMI"):
        """우클릭 컨텍스트 메뉴 표시 및 현재 모드에 따라 명령 구성"""
        tree = self.paut_preview_tree if mode == "PAUT" else (self.rt_preview_tree if mode == "RT" else (self.pt_preview_tree if mode == "PT" else self.preview_tree))
        col = tree.identify_column(event.x)
        self.last_clicked_col = col 
        item_id = tree.identify_row(event.y)
        
        # 선택 상태 보정 (우클릭한 항목으로 선택 변경)
        if item_id:
            if item_id not in tree.selection():
                tree.selection_set(item_id)
            tree.focus(item_id)

        # 메뉴 명령 동적 업데이트
        self.ctx_menu.entryconfigure("선택 항목 체크/해제 토글 (Toggle Check)", command=lambda: self.toggle_selected_items(mode))
        self.ctx_menu.entryconfigure("현재 셀 내용만 복사 (Copy Cell)", command=lambda: self.copy_cell(mode, target="cell"))
        self.ctx_menu.entryconfigure("선택 행 전체 복사 (Copy Row)", command=lambda: self.copy_cell(mode, target="row"))
        self.ctx_menu.entryconfigure("이 컬럼 전체 복사 (Copy Column)", command=lambda: self.copy_cell(mode, target="column"))
        self.ctx_menu.entryconfigure("셀 내용 붙여넣기 (Paste)", command=lambda: self.paste_cell(mode))
        
        self.ctx_menu.entryconfigure("위로 이동 (Move Up)", command=lambda: self.move_item(-1, mode))
        self.ctx_menu.entryconfigure("아래로 이동 (Move Down)", command=lambda: self.move_item(1, mode))
        
        self.ctx_menu.entryconfigure("행 추가 (Add Row)", command=lambda: self.add_item(mode))
        self.ctx_menu.entryconfigure("선택 삭제 (Delete Selected)", command=lambda: self.delete_item(mode))
        
        # PMI 전용/특화 기능 제어
        is_pmi = (mode == "PMI")
        self.ctx_menu.entryconfigure("선택 항목 ISO 병합 (Merge ISO)", command=lambda: self.merge_selected_iso(mode), state="normal" if is_pmi else "disabled")
        self.ctx_menu.entryconfigure("선택 항목 Joint 병합 (Merge Joint)", command=lambda: self.merge_selected_joint(mode), state="normal" if is_pmi else "disabled")
        self.ctx_menu.entryconfigure("선택 항목 ISO 병합 해제 (Ungroup ISO)", command=lambda: self.ungroup_selected_iso(mode), state="normal" if is_pmi else "disabled")
        self.ctx_menu.entryconfigure("선택 항목 Joint 병합 해제 (Ungroup Joint)", command=lambda: self.ungroup_selected_joint(mode), state="normal" if is_pmi else "disabled")
        self.ctx_menu.entryconfigure("선택 항목 일괄 변경 (Bulk Update)", command=lambda: self.show_bulk_update_dialog(mode))

        # 컬럼 설정 기능 (모든 모드 지원)
        self.ctx_menu.entryconfigure("컬럼 설정 (Column Manager)", command=lambda: self.manage_columns(mode))
        self.ctx_menu.entryconfigure("컬럼 설정 (Column Manager)", state="normal")

        self.ctx_menu.post(event.x_root, event.y_root)

    def manage_columns(self, mode="PMI"):
        """프로그램의 각 모드별 미리보기 컬럼 보이기/숨기기 관리"""
        if mode == "RT":
            tree, idx_map, data, keys_attr = self.rt_preview_tree, self.rt_item_idx_map, self.rt_extracted_data, "rt_column_keys"
        elif mode == "KOGAS":
            tree, idx_map, data, keys_attr = self.kogas_preview_tree, self.kogas_item_idx_map, self.kogas_extracted_data, "kogas_column_keys"
        elif mode == "PT":
            tree, idx_map, data, keys_attr = self.pt_preview_tree, self.pt_item_idx_map, self.pt_extracted_data, "pt_column_keys"
        elif mode == "PAUT":
            tree, idx_map, data, keys_attr = self.paut_preview_tree, self.paut_item_idx_map, self.paut_extracted_data, "paut_column_keys"
        else:
            tree, idx_map, data, keys_attr = self.preview_tree, self.item_idx_map, self.extracted_data, "column_keys"

        dialog = tk.Toplevel(self.root)
        dialog.title(f"{mode} 컬럼 관리")
        dialog.geometry("400x550")
        dialog.transient(self.root)
        dialog.grab_set()
        
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill='both', expand=True)
        
        ttk.Label(main_frame, text=f"{mode} 표시할 컬럼을 선택하세요:", font=("Malgun Gothic", 10, "bold")).pack(pady=(0, 10), anchor='w')
        
        # 엑셀 데이터 또는 기존 정의에서 모든 키 추출
        all_keys = []
        if data:
            all_keys = list(data[0].keys())
        else:
            # 데이터가 없을 경우 현재 정의된 키들을 기본으로 함
            all_keys = list(getattr(self, keys_attr))
            
        # 내부 관리용 키 제외
        exclude = ["selected", "date_filtered", "order_index", "visual_group_joint", "is_merged_iso", "is_merged_joint"]
        display_keys = [k for k in all_keys if k not in exclude]
        
        # 스크롤 가능한 영역
        scroll_frame = ttk.Frame(main_frame)
        scroll_frame.pack(fill='both', expand=True)
        
        canvas = tk.Canvas(scroll_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(scroll_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Mouse wheel support for popup
        def _on_popup_mw(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_popup_mw)
        
        # RT용 레이블 매핑 정의 (표준화된 ID 대응)
        label_map = {}
        if mode == "RT" or mode == "KOGAS":
            label_map = {
                "Dwg": "Drawing No.", "Joint": "Film Ident. No.", "Loc": "Film Location",
                "Acc": "Acc", "Rej": "Rej", "Deg": "Deg", "Welder": "Welder No",
                "T": "T", "Mat": "Mat", "No": "No", "Date": "Date", "Remarks": "Remarks",
                "D1": "① Crack", "D2": "② IP", "D3": "③ LF", "D4": "④ Slag", "D5": "⑤ Por",
                "D6": "⑥ U/C", "D7": "⑦ RUC", "D8": "⑧ BT", "D9": "⑨ TI", "D10": "⑩ CP",
                "D11": "⑪ RC", "D12": "⑫ Mis", "D13": "⑬ EP", "D14": "⑭ SD", "D15": "⑮ Oth"
            }
        elif mode == "PAUT":
            label_map = {"ISO": "ISO/DWG", "Location": "Loc"}

        vars_dict = {}
        current_keys = list(getattr(self, keys_attr))
        
        # [ORDERING] 표시할 키들을 현재 설정된 순서대로 정렬 (나머지는 뒤로)
        ordered_keys = [k for k in current_keys if k in display_keys]
        for k in display_keys:
            if k not in ordered_keys: ordered_keys.append(k)
            
        def _refresh_list():
            for widget in scrollable_frame.winfo_children():
                widget.destroy()
            
            for i, k in enumerate(ordered_keys):
                f = tk.Frame(scrollable_frame, background="#ffffff")
                f.pack(fill='x', pady=2)
                
                v = vars_dict.get(k)
                if v is None:
                    v = tk.BooleanVar(value=(k in current_keys))
                    vars_dict[k] = v
                
                display_name = label_map.get(k, k)
                cb = ttk.Checkbutton(f, text=display_name, variable=v)
                cb.pack(side='left', padx=5)
                
                # Up/Down Buttons
                btn_f = tk.Frame(f, background="#ffffff")
                btn_f.pack(side='right', padx=5)
                
                if i > 0:
                    ttk.Button(btn_f, text="▲", width=2, command=lambda _i=i: _move(_i, -1)).pack(side='left')
                if i < len(ordered_keys) - 1:
                    ttk.Button(btn_f, text="▼", width=2, command=lambda _i=i: _move(_i, 1)).pack(side='left')

        def _move(idx, delta):
            target = idx + delta
            ordered_keys[idx], ordered_keys[target] = ordered_keys[target], ordered_keys[idx]
            _refresh_list()

        _refresh_list()

        # 컬럼 관리 버튼 영역 (추가/삭제)
        mgr_btn_frame = ttk.Frame(main_frame)
        mgr_btn_frame.pack(fill='x', pady=5)
        
        from tkinter import simpledialog

        def _add_col():
            name = simpledialog.askstring("새 컬럼 추가", "추가할 컬럼 이름을 입력하세요:", parent=dialog)
            if name:
                name = name.strip()
                if not name: return
                if any(k.lower() == name.lower() for k in all_keys):
                    messagebox.showerror("오류", "이미 존재하는 컬럼 이름입니다.", parent=dialog)
                    return
                # 데이터 구조에 반영
                if data:
                    for item in data:
                        if name not in item: item[name] = ""
                # 현재 키 리스트에도 즉시 반영 (데이터가 없을 때를 위해)
                curr = getattr(self, keys_attr)
                if name not in curr: curr.append(name)

                self.log(f"➕ {mode} 새 컬럼 추가됨: {name}")
                dialog.destroy()
                self.manage_columns(mode)

        def _del_col():
            name = simpledialog.askstring("컬럼 삭제", "삭제할 컬럼 이름을 정확히 입력하세요:", parent=dialog)
            if name:
                name = name.strip()
                if not name: return
                essential = ["V", "No", "selected", "ISO", "Joint", "ISO/DWG", "Dwg"]
                if any(k.lower() == name.lower() for k in essential):
                    messagebox.showerror("오류", f"'{name}' 컬럼은 필수 항목이므로 삭제할 수 없습니다.", parent=dialog)
                    return
                
                target_key = next((k for k in display_keys if k.lower() == name.lower()), None)
                if not target_key:
                    messagebox.showerror("오류", f"'{name}' 컬럼을 찾을 수 없습니다.", parent=dialog)
                    return
                
                if messagebox.askyesno("삭제 확인", f"'{target_key}' 컬럼을 데이터에서 완전히 삭제하시겠습니까?", parent=dialog):
                    if data:
                        for item in data:
                            if target_key in item: del item[target_key]
                    curr = getattr(self, keys_attr)
                    if target_key in curr: curr.remove(target_key)

                    self.log(f"🗑️ {mode} 컬럼 삭제됨: {target_key}")
                    dialog.destroy()
                    self.manage_columns(mode)

        ttk.Button(mgr_btn_frame, text="컬럼 추가", command=_add_col).pack(side='left', padx=5, expand=True)
        ttk.Button(mgr_btn_frame, text="컬럼 삭제", command=_del_col).pack(side='left', padx=5, expand=True)
            
        def _apply():
            selected_keys = []
            
            if "_status" in vars_dict and vars_dict["_status"].get():
                selected_keys.append("_status")
            elif "ST" in vars_dict and vars_dict["ST"].get():
                selected_keys.append("ST")
                
            selected_keys.append("selected") # "V" column is required
            
            if "No" in vars_dict and vars_dict["No"].get():
                selected_keys.append("No")
            elif "No" not in selected_keys:
                selected_keys.append("No") # Fallback safety

            # [FIX] 사용자가 조정한 순서(ordered_keys)를 기준으로 컬럼 수집
            for k in ordered_keys:
                if k not in ["ST", "_status", "No", "selected"] and vars_dict[k].get():
                    selected_keys.append(k)
            
            setattr(self, keys_attr, selected_keys)
            
            # [FIX] RT/KOGAS 모드는 전용 UI 갱신 함수 호출로 처리 (ID 정합성 보장)
            if mode in ["RT", "KOGAS"]:
                self._update_rt_preview_columns(mode=mode)
                self._refresh_rt_config_tabs() # [NEW] 세부 설정 탭도 함께 새로고침
            else:
                # 타 모드(PMI 등)용 기존 로직
                header_mapping = {}
                current_cols = []
                old_display_cols = list(tree["displaycolumns"])
                if old_display_cols == ["#all"] or not old_display_cols:
                    old_display_cols = list(tree["columns"])
                
                for k in selected_keys:
                    col_id = k
                    if k == "selected": col_id = "V"
                    elif k == "_status": col_id = "ST"
                    elif mode == "PMI":
                        if k == "Dwg": col_id = "Dwg"
                        elif k == "Joint": col_id = "Joint"
                    
                    current_cols.append(col_id)
                    header_mapping[col_id] = label_map.get(k, col_id)

                tree["columns"] = tuple(current_cols)
                new_display = [x for x in old_display_cols if x in current_cols]
                for x in current_cols:
                    if x not in new_display: new_display.append(x)
                tree["displaycolumns"] = tuple(new_display)
                
                for col in tree["columns"]:
                    w = 80
                    tree.heading(col, text=header_mapping.get(col, col), anchor='center', command=lambda _c=col: self.sort_by_column(_c, mode=mode))
                    tree.column(col, width=w, anchor='center', stretch=False)
            
            self.populate_preview(data if data else [], switch_tab=False, mode=mode)
            self.log(f"⚙️ {mode} 미리보기 컬럼 설정 변경 완료")
            dialog.destroy()
            
        ttk.Button(main_frame, text="적용 (Apply)", command=_apply).pack(pady=5)


    def update_date_listbox(self, mode="PMI"):
        if mode == "RT":
            listbox = self.rt_date_listbox
            data = self.rt_extracted_data
        elif mode == "KOGAS":
            listbox = self.kogas_date_listbox
            data = self.kogas_extracted_data
        elif mode == "PT":
            listbox = self.pt_date_listbox
            data = self.pt_extracted_data
        else:
            listbox = self.date_listbox
            data = self.extracted_data
        
        if not listbox: return
        
        listbox.delete(0, tk.END)
        # [NEW] 데이터 기반으로 날짜별 선택 상태(date_filtered) 수집
        date_status = {}
        for item in data:
            dt = item.get('Date', 'N/A')
            if dt not in date_status:
                date_status[dt] = item.get('date_filtered', True)
        
        dates = sorted(list(date_status.keys()))
        for d in dates:
            is_v = date_status.get(d, True)
            prefix = "[v]" if is_v else "[ ]"
            listbox.insert(tk.END, f"{prefix} {d}")
            if is_v: listbox.select_set(tk.END)

    def _get_mode_info(self, mode):
        """Helper to get core UI/Data objects for a specific module mode."""
        if mode == "PMI":
            return (self.preview_tree, self.item_idx_map, self.extracted_data, self.column_keys)
        elif mode == "RT":
            return (self.rt_preview_tree, self.rt_item_idx_map, self.rt_extracted_data, self.rt_column_keys)
        elif mode == "KOGAS":
            return (self.kogas_preview_tree, self.kogas_item_idx_map, self.kogas_extracted_data, self.kogas_column_keys)
        elif mode == "PT":
            return (self.pt_preview_tree, self.pt_item_idx_map, self.pt_extracted_data, self.pt_column_keys)
        elif mode == "PAUT":
            return (self.paut_preview_tree, self.paut_item_idx_map, self.paut_extracted_data, self.paut_column_keys)
        return None

    def copy_cell(self, mode="PMI", target="smart"):
        """선택된 영역의 내용을 엑셀 형식(\t)으로 클립보드에 복사"""
        try:
            mode_info = self._get_mode_info(mode)
            if not mode_info: return
            tree, idx_map, data_list, col_keys = mode_info

            selected = list(tree.selection())
            if not selected: return
            selected.sort(key=lambda x: tree.index(x))

            # [REFINED] 드래그 영역 계산
            s_col = getattr(self, 'start_col', None)
            e_col = getattr(self, 'end_col', None)
            l_col = getattr(self, 'last_clicked_col', None)

            if target == "cell":
                target_col_ids = [l_col or s_col or "#4"]
            elif not s_col or not e_col or s_col == e_col:
                target_col_ids = [l_col or s_col or "#4"]
            else:
                try:
                    s_idx = int(s_col.replace('#', '')) - 1
                    e_idx = int(e_col.replace('#', '')) - 1
                    col_start, col_end = min(s_idx, e_idx), max(s_idx, e_idx)
                    target_col_ids = [f"#{i+1}" for i in range(col_start, col_end + 1)]
                except:
                    target_col_ids = [l_col or s_col or "#4"]

            key_map = {f"#{i+1}": k for i, k in enumerate(col_keys)}
            output = []

            for item_id in selected:
                view_idx = tree.index(item_id)
                if not (0 <= view_idx < len(idx_map)): continue
                actual_idx = idx_map[view_idx]
                item = data_list[actual_idx]
                
                row_vals = []
                for c_id in target_col_ids:
                    key = key_map.get(c_id)
                    if key == "selected":
                        row_vals.append("●" if item.get('selected', True) else "○")
                    elif key == "_status":
                        is_def = False
                        if mode == "PMI" and hasattr(self, 'element_filters'):
                            for f in self.element_filters:
                                fk = f['key'].get().strip()
                                if not fk: continue
                                v = self.to_float(self._get_val_ci(item, fk))
                                fm, fx = self.to_float(f['min'].get()), self.to_float(f['max'].get())
                                if (fm > 0 and v < fm) or (fx > 0 and v > fx):
                                    is_def = True; break
                        row_vals.append("⚠️" if is_def else "✅")
                    else:
                        val = item.get(key, "")
                        if mode == "PMI" and key not in ["No", "Joint", "Loc", "Grade", "Date", "Dwg", "_status", "ST", "V", "selected", "order_index"]:
                            f_val = self.to_float(val)
                            val = f"{f_val:.2f}" if f_val > 0 else ""
                        row_vals.append(str(val))
                
                output.append("\t".join(row_vals))

            self.root.clipboard_clear()
            self.root.clipboard_append("\n".join(output))
            self.log(f"📋 {mode} 복사 완료: {len(selected)}행 x {len(target_col_ids)}열")
        except Exception as e:
            self.log(f"❌ 복사 오류: {e}")

    def paste_cell(self, mode="PMI"):
        """클립보드 내용을 선택된 영역에 스마트하게(Replicate 지원) 붙여넣기"""
        try:
            mode_info = self._get_mode_info(mode)
            if not mode_info: return
            tree, idx_map, data_list, col_keys = mode_info
            
            try: 
                clipboard_val = self.root.clipboard_get()
                if not clipboard_val: return
                # 엑셀 형식(\t, \n) 파싱
                paste_rows = [line.split("\t") for line in clipboard_val.strip().split("\n")]
            except Exception as e:
                self.log(f"⚠️ 클립보드 데이터를 가져올 수 없습니다: {e}")
                return
            
            selected = list(tree.selection())
            if not selected:
                messagebox.showwarning("붙여넣기 오류", "붙여넣을 셀을 선택해주세요.")
                return
            selected.sort(key=lambda x: tree.index(x))
            
            # [SMART AREA] 드래그된 컬럼 범위를 계산
            s_col = getattr(self, 'start_col', None)
            e_col = getattr(self, 'end_col', None)
            l_col = getattr(self, 'last_clicked_col', None)
            
            if not s_col or not e_col or s_col == e_col:
                target_col = l_col or s_col or "#4"
                try:
                    s_idx = int(target_col.replace('#', '')) - 1
                    col_start = s_idx
                    # 단일 셀 선택 시 클립보드 너비만큼 자동 확장
                    max_cols_in_paste = len(paste_rows[0]) if paste_rows else 1
                    col_end = min(len(col_keys) - 1, col_start + max_cols_in_paste - 1)
                except:
                    col_start, col_end = 0, len(col_keys) - 1
            else:
                try:
                    s_idx = int(s_col.replace('#', '')) - 1
                    e_idx = int(e_col.replace('#', '')) - 1
                    col_start, col_end = min(s_idx, e_idx), max(s_idx, e_idx)
                except:
                    col_start, col_end = 0, len(col_keys) - 1
            
            target_col_ids = [f"#{i+1}" for i in range(col_start, col_end + 1)]
            key_map = {f"#{i+1}": k for i, k in enumerate(col_keys)}
            
            # [EXCEL-LIKE REPLICATION]
            # 한 행만 선택했는데 클립보드가 여러 행이면 -> 자동 드래그 (Fill Down)
            all_children = tree.get_children('')
            if len(selected) == 1 and len(paste_rows) > 1:
                curr_idx = all_children.index(selected[0])
                for i in range(1, len(paste_rows)):
                    if curr_idx + i < len(all_children):
                        selected.append(all_children[curr_idx + i])
            
            for r_idx, item_id in enumerate(selected):
                view_idx = tree.index(item_id)
                if not (0 <= view_idx < len(idx_map)): continue
                actual_idx = idx_map[view_idx]
                
                # 행 복제(Replicate): 클립보드 행보다 선택된 행이 많으면 순환하며 붙여넣음
                source_row = paste_rows[r_idx % len(paste_rows)]
                
                for c_idx, c_id in enumerate(target_col_ids):
                    # 열 복제(Replicate): 클립보드 열보다 선택된 열이 많으면 순환하며 붙여넣음
                    cell_val = source_row[c_idx % len(source_row)].strip()
                    key = key_map.get(c_id)
                    
                    if key and key not in ["selected", "_status"]:
                        if mode == "PMI" and key not in ["No", "Joint", "Loc", "Grade", "Date", "Dwg", "_status", "ST", "V", "selected", "order_index"]:
                            f_val = self.to_float(cell_val)
                            data_list[actual_idx][key] = f"{f_val:.2f}" if f_val > 0 else ""
                        else:
                            data_list[actual_idx][key] = cell_val
            
            self.populate_preview(data_list, switch_tab=False, mode=mode)
            self.log(f"📋 {mode} 붙여넣기 완료: {len(selected)}행 x {len(target_col_ids)}열")
        except Exception as e:
            self.log(f"❌ 붙여넣기 오류: {e}")
            messagebox.showerror("붙여넣기 실패", f"데이터를 붙여넣는 중 오류가 발생했습니다:\n{e}")

    def merge_selected_iso(self, mode="PMI"):
        """선택된 항목들의 ISO 번호를 첫 번째 항목의 것으로 통일 (병합 효과)"""
        mode_info = self._get_mode_info(mode)
        if not mode_info: return
        tree, idx_map, data_list, _ = mode_info
        
        selected = tree.selection()
        if len(selected) < 2:
            messagebox.showinfo("알림", "병합할 항목을 2개 이상 선택해주세요.")
            return
        
        selected = sorted(list(selected), key=lambda x: tree.index(x))
        view_idx_0 = tree.index(selected[0])
        actual_idx_0 = idx_map[view_idx_0] if 0 <= view_idx_0 < len(idx_map) else None
        
        k_iso = "ISO" if mode == "PAUT" else "Dwg"
        if actual_idx_0 is not None:
            first_iso = data_list[actual_idx_0].get(k_iso, '')
        else:
            first_iso = ""
        
        for i, item_id in enumerate(selected):
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                data_list[actual_idx][k_iso] = first_iso
                if i > 0: data_list[actual_idx]['is_merged_iso'] = True
                else: data_list[actual_idx].pop('is_merged_iso', None)
        
        self.populate_preview(data_list, switch_tab=False, mode=mode)
        self.log(f"🔗 {mode} {len(selected)}개 항목 ISO 병합 완료: {first_iso}")

    def merge_selected_joint(self, mode="PMI"):
        """선택된 항목들의 Joint No를 첫 번째 항목의 것으로 통일 (데이터 완전 변경)"""
        mode_info = self._get_mode_info(mode)
        if not mode_info: return
        tree, idx_map, data_list, _ = mode_info
        
        selected = tree.selection()
        if len(selected) < 2:
            messagebox.showinfo("알림", "병합할 항목을 2개 이상 선택해주세요.")
            return
        
        selected = sorted(list(selected), key=lambda x: tree.index(x))
        view_idx_0 = tree.index(selected[0])
        actual_idx_0 = idx_map[view_idx_0] if 0 <= view_idx_0 < len(idx_map) else None
        
        if actual_idx_0 is not None:
            first_joint = data_list[actual_idx_0].get('Joint', '')
        else:
            first_joint = ""
        
        for i, item_id in enumerate(selected):
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                data_list[actual_idx]['Joint'] = first_joint
                if i > 0: data_list[actual_idx]['is_merged_joint'] = True
                else: data_list[actual_idx].pop('is_merged_joint', None)
        
        self.populate_preview(data_list, switch_tab=False, mode=mode)
        self.log(f"🔗 {mode} {len(selected)}개 항목 Joint 병합 완료: {first_joint}")

    def group_selected_joint(self, mode="PMI"):
        """선택된 항목들을 시각적으로만 같은 Joint로 취급하여 묶음 (데이터 유지)"""
        tree = self.rt_preview_tree if mode == "RT" else self.preview_tree
        idx_map = self.rt_item_idx_map if mode == "RT" else self.item_idx_map
        data = self.rt_extracted_data if mode == "RT" else self.extracted_data
        
        selected = tree.selection()
        if len(selected) < 2:
            messagebox.showinfo("알림", "그룹화할 항목을 2개 이상 선택해주세요.")
            return
            
        selected = sorted(list(selected), key=lambda x: tree.index(x))
            
        view_idx_0 = tree.index(selected[0])
        actual_idx_0 = idx_map[view_idx_0] if 0 <= view_idx_0 < len(idx_map) else None
        
        if actual_idx_0 is not None:
            group_val = data[actual_idx_0].get('Joint', '')
        else:
            group_val = ""
            
        for i, item_id in enumerate(selected):
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                data[actual_idx]['visual_group_joint'] = group_val
                if i > 0: data[actual_idx]['is_merged_joint'] = True
                else: data[actual_idx].pop('is_merged_joint', None)
                
        self.populate_preview(data, switch_tab=False, mode=mode)
        self.log(f"🔗 {mode} {len(selected)}개 항목 Joint 시각적 그룹화 완료")

    def ungroup_selected_iso(self, mode="PMI"):
        """선택된 항목들의 ISO 시각적 그룹화 연결 끊기"""
        tree = self.rt_preview_tree if mode == "RT" else self.preview_tree
        idx_map = self.rt_item_idx_map if mode == "RT" else self.item_idx_map
        data = self.rt_extracted_data if mode == "RT" else self.extracted_data
        
        selected = tree.selection()
        if not selected: return
            
        for item_id in selected:
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                data[actual_idx].pop('visual_group_iso', None)
                data[actual_idx].pop('is_merged_iso', None)
                
        self.populate_preview(data, switch_tab=False, mode=mode)
        self.log(f"✂️ {mode} {len(selected)}개 항목 ISO 병합 해제 완료")

    def ungroup_selected_iso(self, mode="PMI"):
        """선택된 항목들의 시각적 ISO 병합 해제"""
        tree = self.rt_preview_tree if mode == "RT" else self.preview_tree
        idx_map = self.rt_item_idx_map if mode == "RT" else self.item_idx_map
        data = self.rt_extracted_data if mode == "RT" else self.extracted_data
        
        selected = tree.selection()
        if not selected: return
            
        for item_id in selected:
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                data[actual_idx].pop('visual_group_iso', None)
                data[actual_idx].pop('is_merged_iso', None)
                
        self.populate_preview(data, switch_tab=False, mode=mode)
        self.log(f"✂️ {mode} {len(selected)}개 항목 ISO 병합 해제 완료")

    def ungroup_selected_joint(self, mode="PMI"):
        """선택된 항목들의 Joint 시각적 그룹화 연결 끊기"""
        tree = self.rt_preview_tree if mode == "RT" else self.preview_tree
        idx_map = self.rt_item_idx_map if mode == "RT" else self.item_idx_map
        data = self.rt_extracted_data if mode == "RT" else self.extracted_data
        
        selected = tree.selection()
        if not selected: return
            
        for item_id in selected:
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                data[actual_idx].pop('visual_group_joint', None)
                data[actual_idx].pop('is_merged_joint', None)
                
        self.populate_preview(data, switch_tab=False, mode=mode)
        self.log(f"✂️ {mode} {len(selected)}개 항목 Joint 병합 해제 완료")

    def select_all(self, mode="PMI"):
        """모든 항목 체크"""
        if mode == "RT": data = self.rt_extracted_data
        elif mode == "PT": data = self.pt_extracted_data
        elif mode == "PAUT": data = self.paut_extracted_data
        else: data = self.extracted_data
        
        for item in data:
            if item.get('date_filtered', True):
                item['selected'] = True
        self.populate_preview(data, switch_tab=False, mode=mode)

    def deselect_all(self, mode="PMI"):
        """모든 항목 체크 해제"""
        if mode == "RT": data = self.rt_extracted_data
        elif mode == "PT": data = self.pt_extracted_data
        elif mode == "PAUT": data = self.paut_extracted_data
        else: data = self.extracted_data
        
        for item in data:
            if item.get('date_filtered', True):
                item['selected'] = False
        self.populate_preview(data, switch_tab=False, mode=mode)

    def select_deficient_items(self):
        """PMI: 함량 미달인 항목만 자동으로 선택(체크)"""
        if not self.extracted_data or not hasattr(self, 'element_filters'): return
        
        count = 0
        for item in self.extracted_data:
            is_deficient = False
            for f_item in self.element_filters:
                f_key = f_item['key'].get().strip()
                if not f_key: continue
                f_min = round(self.to_float(f_item['min'].get()), 2)
                f_max = round(self.to_float(f_item['max'].get()), 2)
                
                val = round(self.to_float(self._get_val_ci(item, f_key)), 2)
                if (f_min > 0 and val < f_min) or (f_max > 0 and val > f_max):
                    is_deficient = True
                    break
            
            item['selected'] = is_deficient
            if is_deficient: count += 1
            
        self.populate_preview(self.extracted_data, switch_tab=False, mode="PMI")
        self.log(f"✅ 미달 항목 {count}건을 선택했습니다.")

    def reset_pmi_filters(self, refresh=True, clear_search=True):
        """PMI 검색 및 함량 필터 초기화"""
        if clear_search:
            self.pmi_search_loc.set("")
            # 사이드바 지점 필터도 초기화
            if hasattr(self, 'pmi_loc_listbox'):
                for i in range(self.pmi_loc_listbox.size()):
                    val = self.pmi_loc_listbox.get(i)
                    self.pmi_loc_listbox.delete(i)
                    self.pmi_loc_listbox.insert(i, val.replace("[v] ", "[ ] "))
        
        self.pmi_show_deficiency_only.set(False)
        if hasattr(self, 'element_filters'):
            for f in self.element_filters:
                f['min'].set("")
                f['max'].set("")
        if refresh:
            self.populate_preview(self.extracted_data, switch_tab=False, mode="PMI")
        self.log("🧹 PMI 필터가 초기화되었습니다.")

    def update_pmi_loc_listbox(self):
        """현재 데이터의 Test Location 목록을 추출하여 드롭다운 갱신 (날짜 필터 적용된 항목만)"""
        if not hasattr(self, 'pmi_loc_listbox') or not self.extracted_data: return
        
        # 기존 선택 세트 백업
        selected_locs = set()
        for i in range(self.pmi_loc_listbox.size()):
            val = self.pmi_loc_listbox.get(i)
            if val.startswith("[v] "):
                selected_locs.add(val.replace("[v] ", ""))
        
        # 새 목록 추출
        current_locs = sorted(list(set(str(item.get('Loc', '')).strip() for item in self.extracted_data if item.get('Loc') and item.get('date_filtered', True))))
        
        self.pmi_loc_listbox.delete(0, tk.END)
        for loc in current_locs:
            prefix = "[v] " if loc in selected_locs else "[ ] "
            self.pmi_loc_listbox.insert(tk.END, prefix + loc)

    def add_item(self, mode="PMI"):
        """새 데이터 행 추가"""
        if mode == "RT": data = self.rt_extracted_data
        elif mode == "PT": data = self.pt_extracted_data
        elif mode == "PAUT": data = self.paut_extracted_data
        else: data = self.extracted_data
        
        new_item = {'selected': True, 'date_filtered': True, 'order_index': len(data)}
        # 기본 키 초기화
        keys = self.rt_column_keys if mode == "RT" else (self.pt_column_keys if mode == "PT" else (self.paut_column_keys if mode == "PAUT" else self.column_keys))
        for k in keys:
            if k not in new_item: new_item[k] = ""
            
        data.append(new_item)
        self.populate_preview(data, switch_tab=False, mode=mode)
        self.log(f"➕ {mode} 새 데이터 항목 추가 완료")

    def move_item(self, direction, mode="PMI"):
        """선택된 아이템의 순서를 위/아래로 이동 (필터 상태 대응)"""
        # direction이 문자열인 경우 처리 ("up" -> -1, "down" -> 1)
        if isinstance(direction, str):
            direction = -1 if direction.lower() == "up" else 1
        
        mode_info = self._get_mode_info(mode)
        if not mode_info: return
        tree, idx_map, data, _ = mode_info
        
        selected_ids = tree.selection()
        if not selected_ids: return
        
        view_indices = sorted([tree.index(sid) for sid in selected_ids])
        
        if direction == -1: # 위로
            if view_indices[0] == 0: return
        else: # 아래로
            if view_indices[-1] == len(tree.get_children()) - 1: return
            
        # 데이터 리스트(data)에서 위치 교환
        # Treeview 인덱스를 item_idx_map을 통해 실제 데이터 인덱스로 변환
        actual_indices = [idx_map[vi] for vi in view_indices]
        
        # 이동 방향에 따라 처리 순서 결정
        if direction == -1: # 위로
            # 가장 위 아이템의 위쪽 아이템 찾기 (Treeview상에서)
            target_view_idx = view_indices[0] - 1
            target_actual_idx = idx_map[target_view_idx]
            
            # 선택된 아이템들을 한 칸씩 위로 (target_actual_idx 앞으로)
            items_to_move = [data.pop(ai) for ai in reversed(actual_indices)]
            # target_actual_idx가 팝업으로 인해 밀렸을 수 있으므로 다시 계산
            # 하지만 팝업 시 뒤쪽부터 뺐으므로 target_actual_idx 위치는 유지됨
            for item in items_to_move:
                data.insert(target_actual_idx, item)
        else: # 아래로
            # 가장 아래 아이템의 아래쪽 아이템 찾기
            target_view_idx = view_indices[-1] + 1
            target_actual_idx = idx_map[target_view_idx]
            
            # 선택된 아이템들을 한 칸씩 아래로 (target_actual_idx 뒤로)
            items_to_move = [data.pop(ai) for ai in actual_indices]
            # 팝업으로 인해 인덱스가 당겨졌으므로 target_actual_idx에서 len(items_to_move) 만큼 보정하거나 단순히 insert
            insert_pos = target_actual_idx - len(items_to_move) + 1
            for item in reversed(items_to_move):
                data.insert(insert_pos, item)
                
        self.populate_preview(data, switch_tab=False, mode=mode)
        
        # 선택 상태 복원 (Treeview 아이템이 다시 생성되므로 인덱스로 재선택)
        new_children = tree.get_children()
        for vi in view_indices:
            new_pos = max(0, min(len(new_children)-1, vi + direction))
            tree.selection_add(new_children[new_pos])

    def delete_item(self, mode="PMI"):
        """선택된 아이템 삭제"""
        if mode == "RT":
            tree, idx_map, data = self.rt_preview_tree, self.rt_item_idx_map, self.rt_extracted_data
        elif mode == "KOGAS":
            tree, idx_map, data = self.kogas_preview_tree, self.kogas_item_idx_map, self.kogas_extracted_data
        elif mode == "PT":
            tree, idx_map, data = self.pt_preview_tree, self.pt_item_idx_map, self.pt_extracted_data
        elif mode == "PAUT":
            tree, idx_map, data = self.paut_preview_tree, self.paut_item_idx_map, self.paut_extracted_data
        else:
            tree, idx_map, data = self.preview_tree, self.item_idx_map, self.extracted_data
        
        selected_ids = tree.selection()
        if not selected_ids: return
        
        if not messagebox.askyesno("삭제 확인", f"선택한 {len(selected_ids)}개 항목을 정말 삭제하시겠습니까?"):
            return
            
        view_indices = sorted([tree.index(sid) for sid in selected_ids], reverse=True)
        for vi in view_indices:
            actual_idx = idx_map[vi]
            data.pop(actual_idx)
            
        self.populate_preview(data, switch_tab=False, mode=mode)
        self.update_date_listbox(mode)
        self.log(f"🗑️ {mode} {len(selected_ids)}개 항목 삭제 완료")

    def save_preview_data(self, mode="PMI"):
        """현재 추출/편집된 데이터를 JSON 파일로 저장"""
        if mode == "RT": data = self.rt_extracted_data
        elif mode == "KOGAS": data = self.kogas_extracted_data
        elif mode == "PT": data = self.pt_extracted_data
        elif mode == "PAUT": data = self.paut_extracted_data
        else: data = self.extracted_data
        if not data:
            messagebox.showwarning("알림", f"{mode} 저장할 데이터가 없습니다.")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON Files", "*.json")],
            initialfile=f"{mode}_Data_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            title=f"{mode} 미리보기 데이터 저장"
        )
        
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=4)
                self.log(f"💾 {mode} 데이터 저장 완료: {os.path.basename(file_path)}")
                messagebox.showinfo("완료", "데이터가 성공적으로 저장되었습니다.")
            except Exception as e:
                self.log(f"❌ {mode} 저장 오류: {e}")
                messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다: {e}")

    def load_preview_data(self, mode="PMI"):
        """저장된 JSON 데이터를 불러와서 목록에 채움"""
        file_path = filedialog.askopenfilename(
            filetypes=[("JSON Files", "*.json")],
            title=f"{mode} 미리보기 데이터 불러오기"
        )
        
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    loaded_data = json.load(f)
                    
                if not isinstance(loaded_data, list):
                    raise ValueError("올바른 데이터 형식이 아닙니다.")
                
                if mode == "RT": data = self.rt_extracted_data
                elif mode == "KOGAS": data = self.kogas_extracted_data
                elif mode == "PT": data = self.pt_extracted_data
                elif mode == "PAUT": data = self.paut_extracted_data
                else: data = self.extracted_data
                
                if data and not messagebox.askyesno("확인", f"현재 작업 중인 {mode} 데이터가 있습니다. 불러온 데이터로 덮어쓰시겠습니까?\n(아니오를 선택하면 기존 데이터에 추가됩니다.)"):
                    data.extend(loaded_data)
                else:
                    if mode == "RT": self.rt_extracted_data = loaded_data
                    elif mode == "KOGAS": self.kogas_extracted_data = loaded_data
                    elif mode == "PT": self.pt_extracted_data = loaded_data
                    elif mode == "PAUT": self.paut_extracted_data = loaded_data
                    else: self.extracted_data = loaded_data
                    data = loaded_data
            
                # [FIX] 필수 속성 보정
                for i, item in enumerate(data):
                    if 'date_filtered' not in item: item['date_filtered'] = True
                    if 'selected' not in item: item['selected'] = True
                    if 'order_index' not in item: item['order_index'] = i
                    
                self.update_date_listbox(mode)
                self.populate_preview(data, mode=mode)
                self.log(f"📂 {mode} 데이터 불러오기 완료: {os.path.basename(file_path)}")
                messagebox.showinfo("완료", f"데이터를 성공적으로 불러왔습니다.\n(총 {len(data)} 건)")
            except Exception as e:
                self.log(f"❌ {mode} 불러오기 오류: {e}")
                messagebox.showerror("오류", f"파일을 불러오는 중 오류가 발생했습니다: {e}")

    def clear_all(self, mode="PMI"):
        """모든 데이터 초기화"""
        if messagebox.askyesno("확인", f"모든 {mode} 데이터를 초기화하시겠습니까?"):
            if mode == "RT": self.rt_extracted_data = []
            elif mode == "KOGAS": self.kogas_extracted_data = []
            elif mode == "PT": self.pt_extracted_data = []
            elif mode == "PAUT": self.paut_extracted_data = []
            else: self.extracted_data = []
            self.update_date_listbox(mode)
            self.populate_preview([], mode=mode)
            self.log(f"🧹 모든 {mode} 데이터 초기화 완료")

    def export_to_excel(self, mode="PMI"):
        """현재 미리보기 목록(필터링/선택 반영)을 엑셀 파일로 내보냄 (서식 및 병합 시인성 유지)"""
        mode_info = self._get_mode_info(mode)
        if not mode_info: return
        _, _, data, col_keys = mode_info
        
        if not data:
            messagebox.showwarning("알림", f"{mode} 내보낼 데이터가 없습니다.")
            return

        # 1. 엑셀에 저장할 대상 데이터 필터링
        filter_enabled = self.show_selected_only.get()
        final_list = [d for d in data if d.get('date_filtered', True) and (not filter_enabled or d.get('selected', True))]
        
        if not final_list:
            messagebox.showinfo("알림", f"{mode} 현재 조건에 맞는 데이터가 없습니다.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"{mode}_Export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            title=f"{mode} 서식 있는 엑셀 파일로 내보내기"
        )
        if not file_path: return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
            import openpyxl.utils

            wb = Workbook()
            ws = wb.active
            ws.title = f"{mode}_Preview"

            # [REFINED] Export with consistent headers from mode_info
            if mode == "PMI":
                # PMI는 특수하게 ST, V 헤더 추가
                headers = ["ST", "V", "No", "Date", "ISO/DWG", "Joint No", "Test Location", "Ni", "Cr", "Mo", "Grade"]
            else:
                # RT/PT/PAUT는 Treeview 컬럼 헤더 그대로 사용
                tree = self._get_mode_info(mode)[0]
                headers = [tree.heading(c)['text'] for c in tree['columns']]
            
            export_rows = []
            for item in final_list:
                row = {}
                if mode == "PMI":
                    # Status Icon
                    is_def = False
                    if hasattr(self, 'element_filters'):
                        for f in self.element_filters:
                            fk = f['key'].get().strip()
                            if not fk: continue
                            v = self.to_float(self._get_val_ci(item, fk))
                            fm, fx = self.to_float(f['min'].get()), self.to_float(f['max'].get())
                            if (fm > 0 and v < fm) or (fx > 0 and v > fx):
                                is_def = True; break
                    
                    row["ST"] = "⚠️" if is_def else "✅"
                    row["V"] = "●" if item.get('selected', True) else "○"
                    row["No"] = item.get('No', '')
                    row["Date"] = item.get('Date', '')
                    row["ISO/DWG"] = item.get('Dwg', '')
                    row["Joint No"] = item.get('Joint', '')
                    row["Test Location"] = item.get('Loc', '')
                    row["Ni"] = round(self.to_float(item.get('Ni', 0)), 2)
                    row["Cr"] = round(self.to_float(item.get('Cr', 0)), 2)
                    row["Mo"] = round(self.to_float(item.get('Mo', 0)), 2)
                    row["Grade"] = item.get('Grade', '')
                else:
                    # Generic mapping for other modes
                    # col_keys: ("_status", "selected", "No", ...)
                    for i, h in enumerate(headers):
                        k = col_keys[i] if i < len(col_keys) else None
                        if k == "_status":
                            row[h] = "✅" 
                        elif k == "selected":
                            row[h] = "●" if item.get('selected', True) else "○"
                        elif k:
                            row[h] = item.get(k, "")
                        else:
                            row[h] = ""
                export_rows.append(row)

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            
            for row_idx, row_data in enumerate(export_rows, 2):
                for col_idx, key in enumerate(headers, 1):
                    val = row_data.get(key, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

            if mode == "PMI":
                column_widths = {1: 5, 2: 5, 3: 8, 4: 12, 5: 25, 6: 15, 7: 15, 8: 8, 9: 8, 10: 8, 11: 15}
            else:
                # RT Column Widths
                column_widths = {1: 5, 2: 8, 3: 12, 4: 25, 5: 15, 6: 10}
                for i in range(7, 22): column_widths[i] = 4 # Defects
                column_widths[22] = 8 # Result
                column_widths[23] = 12 # Welder
                column_widths[24] = 20 # Remarks
                
            for col_idx, width in column_widths.items():
                if col_idx <= len(headers):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = width

            wb.save(file_path)
            self.log(f"📊 {mode} 서식 포함 엑셀 내보내기 완료: {os.path.basename(file_path)}")
            messagebox.showinfo("완료", f"화면 형식과 동일하게 엑셀 파일로 저장했습니다.\n(총 {len(export_rows)} 건)")
        except Exception as e:
            self.log(f"❌ {mode} 엑셀 내보내기 오류: {e}")
            messagebox.showerror("오류", f"엑셀 저장 중 오류가 발생했습니다: {e}")
        finally:
            if 'wb' in locals():
                try: wb.close()
                except: pass

    def sort_by_column(self, col, mode="PMI"):
        """특정 컬럼을 기준으로 데이터 전체 정렬 (문자열/숫자 구분 및 헤더 표시)"""
        # 0. 데이터 대상 선정
        data_targets = {
            "PMI": self.extracted_data,
            "RT": self.rt_extracted_data,
            "PT": self.pt_extracted_data,
            "PAUT": self.paut_extracted_data
        }
        target_list = data_targets.get(mode)
        if not target_list: return
        
        # 1. 정렬 기준 컬럼 매핑 (모든 모드 통합)
        key_map = {
            "selected": "selected", "V": "selected", "No": "No", "Date": "Date", 
            "Dwg": "Dwg", "Joint": "Joint", "Loc": "Loc", "Ni": "Ni", "Cr": "Cr", "Mo": "Mo", "Grade": "Grade",
            "NPS": "NPS", "Thk.": "Thk.", "Material": "Material", "WType": "WType", "Result": "Result",
            "ISO/DWG": "ISO", "t": "t", "h": "h", "l": "l", "d": "d", "Location": "Location", "Nature": "Nature",
            "ST": "_status", "ISO Drawing No.": "Dwg", "Joint No": "Joint", "Test Location": "Loc", "Weld Type": "WType",
            "Drawing No.": "Dwg", "Joint No.": "Joint"
        }
        data_key = key_map.get(col)
        if not data_key: return

        # 2. 정렬 방향 결정 (기존 방향과 같으면 토글)
        attr_col = f"{mode.lower()}_sort_col"
        attr_rev = f"{mode.lower()}_sort_rev"
        
        if getattr(self, attr_col, None) == col:
            setattr(self, attr_rev, not getattr(self, attr_rev, False))
        else:
            setattr(self, attr_col, col)
            setattr(self, attr_rev, False)

        sort_rev = getattr(self, attr_rev, False)

        # 3. 데이터 정렬 (지능형 복합 자연어 정렬)
        def get_natural_key(val):
            if val is None: return [(2, "")]
            s_val = str(val).strip().upper()
            if not s_val: return [(2, "")]
            # 숫자와 문자열 세그먼트를 분리하여 (0, 숫자) 또는 (1, 문자열) 튜플 리스트 반환
            return [(0, int(c)) if c.isdigit() else (1, c) 
                    for c in re.split(r'(\d+)', s_val) if c]

        def get_value(item, key):
            if key == "_status" and mode == "PMI":
                is_def = False
                if hasattr(self, 'element_filters'):
                    for f in self.element_filters:
                        fk = f['key'].get().strip()
                        if not fk: continue
                        fm = self.to_float(f['min'].get())
                        fx = self.to_float(f['max'].get())
                        v = self.to_float(self._get_val_ci(item, fk))
                        if (fm > 0 and v < fm) or (fx > 0 and v > fx):
                            is_def = True; break
                return 0 if is_def else 1
            # '_current' 가 넘어오면 클릭된 컬럼(data_key) 값을 반환
            return item.get(data_key if key=="_current" else key, "")

        # 4. 계층적 정렬 수행 (클릭컬럼 -> Dwg -> Joint 순서로 Tie-break)
        try:
            k_dwg = "ISO" if mode == "PAUT" else "Dwg"
            k_joint = "Joint"
            
            sort_key_func = lambda x: (
                get_natural_key(get_value(x, "_current")),
                get_natural_key(get_value(x, k_dwg)),
                get_natural_key(get_value(x, k_joint)),
                x.get('order_index', 0)
            )

            target_list.sort(key=sort_key_func, reverse=sort_rev)
        except Exception as e:
            self.log(f"⚠️ {mode} 정렬 오류: {e}")

        # 5. 헤더 텍스트 갱신 (화살표)
        tree_map = {
            "PMI": self.preview_tree, "RT": self.rt_preview_tree,
            "PT": self.pt_preview_tree, "PAUT": self.paut_preview_tree
        }
        target_tree = tree_map.get(mode)
        if target_tree:
            for c in target_tree["columns"]:
                current_text = target_tree.heading(c).get("text", c)
                clean_text = current_text.replace("▲ ", "").replace("▼ ", "")
                prefix = ""
                if c == col:
                    prefix = "▼ " if sort_rev else "▲ "
                target_tree.heading(c, text=prefix + clean_text)

        # 6. 화면 갱신
        self.populate_preview(target_list, switch_tab=False, mode=mode)
        self.log(f"📊 {mode}: '{col}' 기준 {'내림차순' if sort_rev else '오름차순'} 정렬 완료")



    def populate_preview(self, data_list, switch_tab=True, mode="PMI"):
        """추출된 데이터를 미리보기 표에 채움 (필터 반영 및 그룹 색상 적용)"""
        if mode == "RT":
            tree = self.rt_preview_tree
            self.rt_item_idx_map = []
            idx_map = self.rt_item_idx_map
        elif mode == "PT":
            tree = self.pt_preview_tree
            self.pt_item_idx_map = []
            idx_map = self.pt_item_idx_map
        elif mode == "PAUT":
            tree = self.paut_preview_tree
            self.paut_item_idx_map = []
            idx_map = self.paut_item_idx_map
        elif mode == "KOGAS":
            tree = self.kogas_preview_tree
            self.kogas_item_idx_map = []
            idx_map = self.kogas_item_idx_map
        else:
            tree = self.preview_tree
            self.item_idx_map = []
            idx_map = self.item_idx_map
            # [NEW] Update Loc List
            self.update_pmi_loc_listbox()
            
        # [NEW] Deficiency Detection & Auto-Reset (PMI Only)
            # [REMOVED global check, moved to local loop below]

        for item in tree.get_children():
            tree.delete(item)
        
        filter_enabled = self.show_selected_only.get()
        hidden_by_def_count = 0
        total_matched_search = 0
        local_pmi_deficient_count = 0 # Count within Date/Loc filter
        
        active_locs = set()
        if mode == "PMI" and hasattr(self, 'pmi_loc_listbox'):
            for i in range(self.pmi_loc_listbox.size()):
                val = self.pmi_loc_listbox.get(i)
                if val.startswith("[v] "): active_locs.add(val.replace("[v] ", "").lower())
        
        last_iso = None
        last_joint = None
        current_tag = "group_even"
        
        for idx, item in enumerate(data_list):
            if not item.get('date_filtered', True):
                continue
            
            is_selected = item.get('selected', True)
            if filter_enabled and not is_selected:
                continue
            
            # [NEW] PMI Preview Filters (Multi-Select Test Location & Deficiency)
            selected_locs_names = []
            if mode == "PMI":
                # item Date/Loc filter already passed (Date was line 2608, Loc is here)
                loc_val = str(item.get('Loc', '')).lower()
                if active_locs and loc_val not in active_locs:
                    continue
                
                # Check for deficiency highlighting (AND local count)
                is_deficient = False
                if hasattr(self, 'element_filters'):
                    for f_item in self.element_filters:
                        f_key = f_item['key'].get().strip()
                        if not f_key: continue
                        f_min = round(self.to_float(f_item['min'].get()), 2)
                        f_max = round(self.to_float(f_item['max'].get()), 2)
                        val = round(self.to_float(self._get_val_ci(item, f_key)), 2)
                        if (f_min > 0 and val < f_min) or (f_max > 0 and val > f_max):
                            is_deficient = True; break
                
                if is_deficient: local_pmi_deficient_count += 1
                total_matched_search += 1
                selected_locs_names.append(item.get('Loc', '')) 
                
                if self.pmi_show_deficiency_only.get() and not is_deficient:
                    hidden_by_def_count += 1
                    continue
            else:
                is_deficient = False

            idx_map.append(idx)
            v_mark = "●" if is_selected else "○"
            st_mark = "⚠️" if is_deficient else "✅"
            
            # ISO/DWG 번호가 바뀌면 배경색 태그 교체 (PAUT 'ISO' 대응)
            curr_iso = item.get('Dwg', item.get('ISO', ''))
            norm_iso = self.normalize_iso(curr_iso)
            
            # [NEW] 시각적 병합 (visual_group_joint) 값이 존재하면 해당 값으로 그룹화 계산
            curr_joint = item.get('visual_group_joint', item.get('Joint', ''))
            
            # [FIX] Joint-Aware: Joint가 바뀌었는지도 함께 체크
            is_new_iso = (last_iso is None or self.normalize_iso(last_iso) != norm_iso)
            is_new_joint = (last_joint is None or last_joint != curr_joint)
            
            if is_new_iso:
                current_tag = "group_odd" if current_tag == "group_even" else "group_even"

            # [REFINED] 하이브리드 표시 로직: 블록 시작점 정렬이거나 "새로운 그룹(ISO or Joint)" 시작점일 때 표시
            display_count = len(idx_map) - 1 
            is_block_start = (display_count % 3 == 0)
            
            # ISO나 Joint가 바뀌었거나 3줄 단위 시작이면 표시
            is_show = is_new_iso or is_new_joint or is_block_start

            display_iso = curr_iso if is_show else ""
            display_joint = item.get('Joint', '') if is_show else ""
            
            # [FIX] 명시적으로 병합된 경우라도 '새 그룹 시작'이면 표시 강제
            if item.get('is_merged_iso') and not (is_new_iso or is_new_joint): display_iso = ""
            if item.get('is_merged_joint') and not (is_new_iso or is_new_joint): display_joint = ""

            last_iso = curr_iso
            last_joint = curr_joint
            
            row_vals = []
            if mode == "RT" or mode == "KOGAS":
                keys = self.kogas_column_keys if mode == "KOGAS" else self.rt_column_keys
                for k in keys:
                    if k == "selected": row_vals.append(v_mark)
                    elif k == "Acc":
                        val = item.get("Acc", "")
                        if not val:
                            res = str(item.get("Result", "")).upper()
                            val = "OK" if "ACC" in res or "OK" in res else ""
                        row_vals.append(val)
                    elif k == "Rej":
                        val = item.get("Rej", "")
                        if not val:
                            res = str(item.get("Result", "")).upper()
                            val = "NG" if "REJ" in res or "NG" in res or "RE" in res else ""
                        row_vals.append(val)
                    elif k == "Deg":
                        row_vals.append(str(item.get("Deg", "")).strip())
                    elif mode == "KOGAS" and k in ["Dwg", "Mat", "Welder"]:
                        main_val = str(item.get(k, "")).strip()
                        sub_val = str(item.get(f"{k}_Sub", "")).strip()
                        if sub_val and sub_val != main_val:
                            row_vals.append(f"{main_val} / {sub_val}")
                        else:
                            row_vals.append(main_val)
                    else: 
                        val = str(item.get(k, "")).strip()
                        if (k in ["Dwg", "ISO"] and item.get('is_merged_iso')) or (k == "Joint" and item.get('is_merged_joint')): val = ""
                        row_vals.append(val)
            elif mode == "PT":
                for k in self.pt_column_keys:
                    if k == "selected": row_vals.append(v_mark)
                    else: 
                        val = str(item.get(k, "")).strip()
                        if (k in ["Dwg", "ISO"] and item.get('is_merged_iso')) or (k == "Joint" and item.get('is_merged_joint')): val = ""
                        row_vals.append(val)
            elif mode == "PAUT":
                # PAUT use paut_column_keys
                for k in self.paut_column_keys:
                    if k == "No": row_vals.append(len(idx_map))
                    else: 
                        val = str(item.get(k, "")).strip()
                        if (k in ["Dwg", "ISO"] and item.get('is_merged_iso')) or (k == "Joint" and item.get('is_merged_joint')): val = ""
                        row_vals.append(val)
            else: # PMI
                for k in self.column_keys:
                    if k == "_status": row_vals.append(st_mark)
                    elif k == "selected": row_vals.append(v_mark)
                    elif k in ["Ni", "Cr", "Mo", "Mn"]:
                        row_vals.append(f"{self.to_float(item.get(k, 0)):.2f}")
                    else:
                        val = str(item.get(k, "")).strip()
                        if (k in ["Dwg", "ISO"] and item.get('is_merged_iso')) or (k == "Joint" and item.get('is_merged_joint')): val = ""
                        row_vals.append(val)

            row_tags = [str(idx), current_tag]
            if is_deficient:
                row_tags.append("deficient")
            tree.insert("", "end", values=tuple(row_vals), tags=tuple(row_tags))
            
        # [REACTIVE AUTO-RESET] Trigger if current view is clean but filters are active
        if mode == "PMI":
            has_pmi_filter = self.pmi_show_deficiency_only.get()
            if not has_pmi_filter and hasattr(self, 'element_filters'):
                for f in self.element_filters:
                    if f['min'].get().strip() or f['max'].get().strip():
                        has_pmi_filter = True; break
            
            if local_pmi_deficient_count == 0 and has_pmi_filter:
                self.reset_pmi_filters(refresh=True, clear_search=False)
                return

        if mode == "PMI" and hidden_by_def_count > 0:
            sel_str = ", ".join(list(set(selected_locs_names))) if selected_locs_names else "전체"
            self.log(f"ℹ️ '{sel_str}' 결과 중 {hidden_by_def_count}건이 '함량 미달만 보기' 옵션으로 인해 숨겨져 있습니다.")
            
        # [REMOVED] Legacy tab switching logic (Preview is now permanently visible in splitscreen)
        # if switch_tab:
        #     if mode == "RT":
        #         self.rt_tab_notebook.select(self.rt_tab_preview)
        #     elif mode == "PT":
        #         self.pt_tab_notebook.select(self.pt_tab_preview)
        #     elif mode == "PAUT":
        #         pass 
        #     else:
        #         self.tab_notebook.select(self.tab_preview)

    def _browse_dir(self, var):
        current = var.get()
        init_dir = current if current and os.path.exists(current) else RESOURCE_DIR
        path = filedialog.askdirectory(initialdir=init_dir)
        if path: var.set(path)

    def _update_file_info(self, mode, path):
        """Extracts and formats file metadata for the UI header."""
        if not path or not os.path.exists(path):
            self.file_info_vars[mode].set("📄 파일을 선택해주세요.")
            return

        try:
            fname = os.path.basename(path)
            fsize = os.path.getsize(path) / 1024 # KB
            mtime = os.path.getmtime(path)
            dt_str = datetime.datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M')
            
            size_str = f"{fsize:.1f}KB" if fsize < 1024 else f"{fsize/1024:.2f}MB"
            self.file_info_vars[mode].set(f"📄 파일: {fname}  |  📏 크기: {size_str}  |  📅 수정일: {dt_str}")
        except Exception as e:
            self.file_info_vars[mode].set(f"⚠️ 파일 정보 읽기 오류: {str(e)}")

    def _sync_all_file_infos(self):
        """Updates all mode headers from current config paths."""
        modes = [
            ('PMI', self.target_file_path.get()),
            ('RT', self.rt_target_file_path.get()),
            ('KOGAS', self.kogas_target_file_path.get()),
            ('PT', self.pt_target_file_path.get()),
            ('PAUT', self.paut_target_file_path.get())
        ]
        for mode, path in modes:
            self._update_file_info(mode, path)

    def _browse_file(self, var, types):
        current = var.get()
        init_dir = os.path.dirname(current) if current and os.path.exists(current) else BASE_DIR
        path = filedialog.askopenfilename(initialdir=init_dir, filetypes=types)
        if path: 
            var.set(path)
            # Find which mode this belongs to and update info
            for mode, m_var in [('PMI', self.target_file_path), ('RT', self.rt_target_file_path), 
                               ('KOGAS', self.kogas_target_file_path), ('PT', self.pt_target_file_path), 
                               ('PAUT', self.paut_target_file_path)]:
                if var == m_var:
                    self._update_file_info(mode, path)
                    break 

    def show_bulk_update_dialog(self, mode="PMI"):
        """선택된 항목들을 필터 조건에 따라 일괄 변경하는 다이얼로그 표시"""
        target_data = self.rt_extracted_data if mode == "RT" else (self.pt_extracted_data if mode == "PT" else self.extracted_data)
        
        if not target_data:
            messagebox.showwarning("알림", f"{mode} 데이터가 없습니다.")
            return

        # 체크(●)된 항목의 인덱스 추출
        selected_indices = [idx for idx, item in enumerate(target_data) if item.get('selected', True)]
        if not selected_indices:
            messagebox.showwarning("항목 미선택", "일괄 변경할 항목을 체크(선택)해주세요.")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title(f"{mode} 선택 항목 일괄 변경")
        dialog.geometry("450x570")
        dialog.configure(background="#f9fafb")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.geometry(f"+{self.root.winfo_x() + 100}+{self.root.winfo_y() + 50}")

        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill='both', expand=True)

        ttk.Label(main_frame, text="1. 대상 필터링 (선택 사항)", font=("Malgun Gothic", 10, "bold")).pack(anchor='w', pady=(0, 10))
        
        filter_frame = ttk.Frame(main_frame)
        filter_frame.pack(fill='x', pady=(0, 10))
        
        # 모드별 필터 가능 열 설정
        if mode == "RT":
            filter_cols = ["전체(필터 없음)", "Drawing No.", "Film Ident. No.", "Film Location", "Welder No", "Remarks"]
            key_map = {"Drawing No.": "Dwg", "Film Ident. No.": "Joint", "Film Location": "Loc", "Welder No": "Welder", "Remarks": "Remarks"}
        elif mode == "PT":
            filter_cols = ["전체(필터 없음)", "ISO Drawing No.", "Joint", "NPS", "Thk.", "Material", "Welder"]
            key_map = {"ISO Drawing No.": "Dwg", "Joint": "Joint", "NPS": "NPS", "Thk.": "Thk.", "Material": "Material", "Welder": "Welder"}
        else: # PMI
            filter_cols = ["전체(필터 없음)", "ISO/DWG", "Joint No", "Test Location", "Ni", "Cr", "Mo", "Grade"]
            key_map = {"ISO/DWG": "Dwg", "Joint No": "Joint", "Test Location": "Loc", "Ni": "Ni", "Cr": "Cr", "Mo": "Mo", "Grade": "Grade"}

        ttk.Label(filter_frame, text="필터 열:").grid(row=0, column=0, sticky='w')
        col_var = tk.StringVar(value="전체(필터 없음)")
        col_combo = ttk.Combobox(filter_frame, textvariable=col_var, state='readonly', width=20, values=filter_cols)
        col_combo.grid(row=0, column=1, padx=5, sticky='w')
        
        ttk.Label(filter_frame, text="필터 값:").grid(row=1, column=0, sticky='w', pady=5)
        val_var = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=val_var, width=23).grid(row=1, column=1, padx=5, sticky='w', pady=5)
        
        # [NEW] Filter Integration Checkboxes
        only_def_var = tk.BooleanVar(value=False)
        only_visible_var = tk.BooleanVar(value=True) # Default to True for convenience
        
        filter_options_frame = ttk.Frame(main_frame)
        filter_options_frame.pack(fill='x', pady=(0, 10))
        
        if mode == "PMI":
            tk.Checkbutton(filter_options_frame, text="⚠️ 함량 미달인 항목만 변경 적용", variable=only_def_var, 
                           background="#f9fafb", font=("Malgun Gothic", 9)).pack(anchor='w')
        
        tk.Checkbutton(filter_options_frame, text="🔍 현재 화면에 보이는(필터링된) 항목만 변경", variable=only_visible_var, 
                       background="#f9fafb", font=("Malgun Gothic", 9)).pack(anchor='w')
        
        ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=5)

        ttk.Label(main_frame, text="2. 변경할 값 입력 (입력된 항목만 반영)", font=("Malgun Gothic", 10, "bold")).pack(anchor='w', pady=(0, 10))
        
        fields_frame = ttk.Frame(main_frame)
        fields_frame.pack(fill='both', expand=True)
        
        # 모드별 입력 필드 구성
        if mode == "RT":
            labels = [("Dwg:", "Dwg"), ("Film Ident:", "Joint"), ("Loc:", "Loc"), ("Grade:", "Grade"), ("Welder:", "Welder"), ("Remarks:", "Remarks"), ("Date:", "Date")]
        elif mode == "PT":
            labels = [("Dwg:", "Dwg"), ("Joint:", "Joint"), ("NPS:", "NPS"), ("Thk:", "Thk."), ("Material:", "Material"), ("Welder:", "Welder"), ("Weld Type:", "WType"), ("Result:", "Result"), ("Date:", "Date")]
        else: # PMI
            labels = [("Ni (%):", "Ni"), ("Cr (%):", "Cr"), ("Mo (%):", "Mo"), ("Location:", "Loc"), ("Grade:", "Grade"), ("Date:", "Date")]

        entries = {}
        for i, (lbl, key) in enumerate(labels):
            ttk.Label(fields_frame, text=lbl).grid(row=i, column=0, sticky='e', pady=3, padx=5)
            entries[key] = ttk.Entry(fields_frame, width=20)
            entries[key].grid(row=i, column=1, sticky='w', pady=3)

        def apply_action():
            filter_col = col_var.get()
            filter_val = val_var.get().strip().lower()
            target_key = key_map.get(filter_col)
            
            update_spec = {}
            for key, entry in entries.items():
                val = entry.get().strip()
                if val:
                    if mode == "PMI" and key not in ["No", "Joint", "Loc", "Grade", "Date", "Dwg", "_status", "ST", "V", "selected", "order_index"]:
                        if val.startswith(('+', '-')):
                            try: update_spec[key] = ('relative', float(val))
                            except: update_spec[key] = ('absolute', self.to_float(val))
                        else:
                            update_spec[key] = ('absolute', self.to_float(val))
                    else:
                        update_spec[key] = ('absolute', val)
            
            if not update_spec:
                messagebox.showwarning("입력 부족", "변경할 값을 하나 이상 입력해주세요.")
                return

            count = 0
            for idx in selected_indices:
                item = target_data[idx]
                match = True
                if target_key:
                    curr_val = str(item.get(target_key, "")).strip().lower()
                    if filter_val not in curr_val: match = False
                
                # [NEW] Check for deficiency if requested
                if match and mode == "PMI" and only_def_var.get():
                    is_deficient = False
                    for f_item in self.element_filters:
                        f_key = f_item['key'].get().strip()
                        if not f_key: continue
                        f_min = round(self.to_float(f_item['min'].get()), 2)
                        f_max = round(self.to_float(f_item['max'].get()), 2)
                        val = round(self.to_float(self._get_val_ci(item, f_key)), 2)
                        if (f_min > 0 and val < f_min) or (f_max > 0 and val > f_max):
                            is_deficient = True; break
                    if not is_deficient: match = False
                
                # [NEW] Check for sidebar visibility (Date & Loc)
                if match and only_visible_var.get():
                    if not item.get('date_filtered', True):
                        match = False
                    
                    if match and mode == "PMI" and hasattr(self, 'pmi_loc_listbox'):
                        # Get active sidebar locs
                        active_locs = set()
                        for i in range(self.pmi_loc_listbox.size()):
                            v = self.pmi_loc_listbox.get(i)
                            if v.startswith("[v] "): active_locs.add(v.replace("[v] ", "").lower())
                        
                        if active_locs:
                            loc_val = str(item.get('Loc', '')).lower()
                            if loc_val not in active_locs:
                                match = False

                if match:
                    for k, spec in update_spec.items():
                        up_mode, up_val = spec
                        if up_mode == 'relative':
                            item[k] = self.to_float(item.get(k, 0)) + up_val
                        else:
                            item[k] = up_val
                    
                    # [NEW] Re-calculate Grade ONLY if NOT manually provided
                    if "Grade" not in update_spec:
                        new_grade = self.check_material_grade(item)
                        if new_grade:
                            item['Grade'] = new_grade
                    
                    count += 1
            
            self.populate_preview(target_data, switch_tab=False, mode=mode)
            messagebox.showinfo("일괄 변경 완료", f"총 {count}개의 항목이 업데이트되었습니다.")
            dialog.destroy()

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=(20, 0))
        ttk.Button(btn_frame, text="적용 (Apply)", command=apply_action).pack(side='right', padx=5)
        ttk.Button(btn_frame, text="취소 (Cancel)", command=dialog.destroy).pack(side='right')

    # --- Integrated Verification Logic ---

    def to_float(self, val):
        try:
            s = str(val).replace('%', '').strip()
            if not s or s.lower() == 'nan': return 0.0
            return float(s)
        except: return 0.0

    def _get_val_ci(self, item, key):
        """대소문자 구분 없이 딕셔너리에서 값 찾기 (Cr vs CR 등)"""
        if not key or not isinstance(item, dict): return None
        if key in item: return item[key]
        k_lower = key.lower()
        for k in item.keys():
            if str(k).lower() == k_lower:
                return item[k]
        return None

    def check_material_grade(self, row_data):
        """jjchRFIPMI.py에서 이식된 10% 여유치 판정 로직"""
        cr = self.to_float(self._get_val_ci(row_data, 'Cr'))
        ni = self.to_float(self._get_val_ci(row_data, 'Ni'))
        mo = self.to_float(self._get_val_ci(row_data, 'Mo'))
        mn = self.to_float(self._get_val_ci(row_data, 'Mn'))
        
        margin = 0.1 # 10% 여유
        
        # 1. SUS 316 (Cr:16~18 / Ni:10~14 / Mo:2~3)
        if (16.0*(1-margin) <= cr <= 18.0*(1+margin)) and (10.0*(1-margin) <= ni <= 14.0*(1+margin)) and (2.0*(1-margin) <= mo <= 3.0*(1+margin)):
            return "SS316"
        
        # 2. Duplex (Cr:22~23 / Mo:3~3.5 / Ni:4.5~6.5 / Mn:2.0이하)
        if (22.0*(1-margin) <= cr <= 23.0*(1+margin)) and (4.5*(1-margin) <= ni <= 6.5*(1+margin)) and (3.0*(1-margin) <= mo <= 3.5*(1+margin)) and (mn <= 2.2):
            return "DUPLEX"

        # 3. SUS 310 (Cr:24~26 / Ni:19~22)
        if (24.0*(1-margin) <= cr <= 26.0*(1+margin)) and (19.0*(1-margin) <= ni <= 22.0*(1+margin)):
            return "SS310"

        # 4. SUS 304 (Cr:18↑ / Ni:8↑ / Mo:0.5↓ / Mn:2.0↓)
        if (cr >= 16.2) and (ni >= 7.2) and (mo <= 0.55) and (mn <= 2.2):
            return "SS304"

        return None # 판정 불가 시 원본 사용

    def normalize_iso(self, val):
        """정렬 및 그룹화 시와 동일하게 정규화된 값으로 비교하여 도면번호 일관성 유지"""
        s_val = str(val).strip()
        try:
            if re.match(r'^\d+(\.\d+)?$', s_val):
                f_val = float(s_val)
                if f_val == int(f_val): return str(int(f_val))
                return str(f_val)
        except: pass
        return s_val.lower()

    # --- Excel Helper Logic ---

    def find_image_smart(self, keyword, exclude_keyword=None):
        def _search_in_folder(folder_path):
            if not folder_path or not os.path.exists(folder_path): return None
            candidates = glob.glob(os.path.join(folder_path, "*.*"))
            valid_extensions = ['.PNG', '.JPG', '.JPEG', '.BMP', '.GIF']
            for path in candidates:
                fname = os.path.basename(path).upper(); ext = os.path.splitext(path)[1].upper()
                if ext not in valid_extensions: continue 
                if keyword.upper() in fname:
                    if exclude_keyword and exclude_keyword.upper() in fname: continue
                    return path 
            return None

        # 1. UI에서 설정한 폴더에서 먼저 검색
        found = _search_in_folder(self.logo_folder_path.get())
        if found: return found
        
        # 2. PyInstaller 묶음(실행파일 내부 임시 폴더)에서 검색 (Standalone 지원)
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            found = _search_in_folder(sys._MEIPASS)
            if found: return found
            
        return None

    def place_image_freely(self, ws, img_path, anchor_cell_str, w, h, x_offset, y_offset):
        try:
            if not img_path or not os.path.exists(img_path): return
            
            # [ENHANCED] Robust image processing with terminal & UI logging
            original = PILImage.open(img_path).convert("RGBA")
            w_px = max(5, int(float(w))); h_px = max(5, int(float(h)))
            
            temp_name = f"temp_{int(time.time())}_{os.path.basename(img_path)}.png"
            temp_full_path = os.path.join(tempfile.gettempdir(), temp_name)
            
            resized = original.resize((w_px, h_px), PILImage.Resampling.LANCZOS)
            resized.save(temp_full_path, "PNG")
            
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(temp_full_path)
            img.width = w_px; img.height = h_px
            from openpyxl.utils import column_index_from_string
            from openpyxl.utils.cell import coordinate_from_string
            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker, XDRPositiveSize2D
            
            # [FIX] Handle 'None' or empty anchor string
            if not anchor_cell_str or str(anchor_cell_str).strip() == "None" or len(str(anchor_cell_str).strip()) < 2:
                anchor_cell_str = "A1"

            try:
                col_str, row_num = coordinate_from_string(anchor_cell_str)
                col_idx = max(0, column_index_from_string(col_str) - 1)
                row_idx = max(0, row_num - 1) 
            except Exception as e:
                col_idx, row_idx = 0, 0
                anchor_cell_str = "A1"
            
            # [STABLE] Use OneCellAnchor instead of AbsoluteAnchor for better Excel compatibility (prevents corruption)
            # 1 Pixel = 9525 EMU (Standard 96 DPI)
            final_emu_x = max(0, int(float(x_offset) * 9525))
            final_emu_y = max(0, int(float(y_offset) * 9525))
            emu_w = max(9525, int(float(w_px) * 9525))
            emu_h = max(9525, int(float(h_px) * 9525))
            
            # [FIX] Use AnchorMarker (cell + offset) for stable positioning
            marker = AnchorMarker(col=col_idx, colOff=final_emu_x, row=row_idx, rowOff=final_emu_y)
            size = XDRPositiveSize2D(cx=emu_w, cy=emu_h)
            
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(img)
            
            msg = f"[SUCCESS] 로고 배치 완료: {os.path.basename(img_path)}"
            self.log(f"   {msg}")
        except Exception as e:
            msg = f"[ERROR] 로고 배치 실패: {e}"
            self.log(f"   {msg}")

    def add_logos_to_sheet(self, ws, is_cover=False, clear_existing=True, mode=None):
        # [SUPER SAFE] Identify if this is the first sheet (Cover/Gapji) of the workbook
        try:
            is_first_sheet = (ws == ws.parent.worksheets[0])
        except:
            is_first_sheet = is_cover

        # [FIX] Never clear existing images on the Cover sheet (Gapji) or in RT mode to preserve template diagrams (e.g. Shooting Sketches)
        if is_first_sheet or is_cover or mode == "RT":
            clear_existing = False

        # [NEW] If RT mode and images already exist, skip adding logos to preserve the template perfectly
        if mode == "RT" and getattr(ws, '_images', None):
            self.log(f"   ℹ️ [RT] 시트에 이미 그림(도면)이 있어 추가 로고 삽입을 건너뜁니다.")
            return

        if clear_existing:
            try: ws._images = [] 
            except: pass
        
        # [ENHANCED] Use provided mode or tracked current_mode
        if mode is None:
            mode = getattr(self, 'current_mode', "PMI")
            try:
                tab_text = self.mode_notebook.tab(self.mode_notebook.select(), "text")
                if "PMI" in tab_text: mode = "PMI"
                elif "RT" in tab_text: mode = "RT"
                elif "PT" in tab_text: mode = "PT"
                elif "PAUT" in tab_text: mode = "PAUT"
            except: pass
            
        ctx = "COVER" if is_cover else "DATA"
        context = f"{mode}_{ctx}"
        
        debug_msg = f"?? 로고 엔진 가동: {context} (Mode: {mode}, Sheet: {ws.title})"
        self.log(debug_msg)
        
        def _get_val(prefix, suffix, default):
            # [FIX] Match ACTUAL UI key format: {prefix}_{mode}_{ctx}{suffix}
            # e.g., SITCO_RT_COVER_X
            full_key = f"{prefix}_{context}{suffix}"
            v = self.config.get(full_key)
            
            # If value is missing or effectively zero for coordinates/size
            is_coord = suffix in ["_X", "_Y", "_W", "_H"]
            if v is not None and v != "":
                try:
                    if is_coord and float(v) <= 0: pass # Fallback if zero
                    else: return v
                except: pass
                
            # [FIX] Enhanced Fallback: If DATA is missing, try COVER of the same mode first
            if "_DATA" in context:
                cover_key = f"{prefix}_" + context.replace("_DATA", "_COVER") + suffix
                cv = self.config.get(cover_key)
                if cv is not None and cv != "":
                    try:
                        if is_coord and float(cv) <= 0: pass
                        else: return cv
                    except: pass

            # [FIX] Enhanced Fallback: If current mode settings are missing, try PMI_COVER
            if mode != "PMI":
                pmi_fallback_key = f"{prefix}_PMI_COVER{suffix}"
                pv = self.config.get(pmi_fallback_key)
                if pv is not None and pv != "":
                    try:
                        if is_coord and float(pv) <= 0: pass
                        else: return pv
                    except: pass

            # [FIX] Final Global Fallback (e.g., SITCO_COVER_X)
            global_ctx = "COVER" if "COVER" in context else "DATA"
            global_key = f"{prefix}_{global_ctx}{suffix}"
            gv = self.config.get(global_key, default)
            if is_coord:
                try:
                    if gv is None or gv == "" or float(gv) <= 0: return default
                except: return default
            return gv

        def _get_effective_path(prefix, keywords):
            # 1. [FIX] Match ACTUAL UI key format: {prefix}_{mode}_{ctx}_PATH
            # e.g., SITCO_RT_COVER_PATH
            k1 = f"{prefix}_{context}_PATH"
            path = self.config.get(k1, "")
            
            # [FIX] Enhanced Fallback: If DATA path is empty, try COVER path of same mode
            if (not path or not str(path).strip()) and "_DATA" in context:
                k2 = f"{prefix}_" + context.replace("_DATA", "_COVER") + "_PATH"
                path = self.config.get(k2, "")
            
            # [FIX] Enhanced Fallback: If current mode path is empty, try PMI_COVER path
            if (not path or not str(path).strip()) and mode != "PMI":
                k3 = f"{prefix}_PMI_COVER_PATH"
                path = self.config.get(k3, "")
            
            # [FIX] Final Global Fallback (e.g., SITCO_COVER_PATH)
            if (not path or not str(path).strip()):
                global_ctx = "COVER" if "COVER" in context else "DATA"
                k4 = f"{prefix}_{global_ctx}_PATH"
                path = self.config.get(k4, "")

            # self.log(f"   > '{prefix}' 검색 시작 (Target Context: {context})")
            
            # 2. If path is valid and exists, use it
            if path and str(path).strip() and os.path.exists(str(path).strip()):
                self.log(f"   [OK] '{prefix}' 수동 지정 경로 발견: {os.path.basename(str(path))}")
                return str(path).strip()
            
            # 3. Fallback to global path if mode-specific is empty
            global_ctx = "COVER" if "COVER" in context else "DATA"
            # [FIX] Match UI key format for global fallback path
            g_path = self.config.get(f"{global_ctx}_{prefix}_PATH", "")
            if g_path and str(g_path).strip() and os.path.exists(str(g_path).strip()):
                self.log(f"   [OK] '{prefix}' 공용 설정 경로 사용: {os.path.basename(str(g_path))}")
                return str(g_path).strip()

            # 4. If manual path is empty or invalid, try Smart Search in Logo Folder
            # self.log(f"   ? '{prefix}' 설정 없음 -> 기본 폴더 자동 검색 시작...")
            for kw in keywords:
                found = self.find_image_smart(kw)
                if found:
                    self.log(f"   [OK] '{prefix}' 자동 검색 성공: {os.path.basename(found)}")
                    return found
            
            # self.log(f"   [FAIL] '{prefix}' 로고를 어디에서도 찾을 수 없습니다.")
            return None

        # 1. SITCO
        p = _get_effective_path("SITCO", ["SITCO"])
        if p: self.place_image_freely(ws, p, _get_val("SITCO", "_ANCHOR", "A1"), float(_get_val("SITCO", "_W", 100)), float(_get_val("SITCO", "_H", 50)), float(_get_val("SITCO", "_X", 0)), float(_get_val("SITCO", "_Y", 0)))
        
        # 2. SEOUL
        p = _get_effective_path("SEOUL", ["서울검사"])
        if p: self.place_image_freely(ws, p, _get_val("SEOUL", "_ANCHOR", "A1"), float(_get_val("SEOUL", "_W", 100)), float(_get_val("SEOUL", "_H", 50)), float(_get_val("SEOUL", "_X", 0)), float(_get_val("SEOUL", "_Y", 0)))
        
        # 3. FOOTER
        p = _get_effective_path("FOOTER", ["바닥글", "PMI"])
        if p: self.place_image_freely(ws, p, _get_val("FOOTER", "_ANCHOR", "A1"), float(_get_val("FOOTER", "_W", 100)), float(_get_val("FOOTER", "_H", 50)), float(_get_val("FOOTER", "_X", 0)), float(_get_val("FOOTER", "_Y", 0)))

        # 4. FOOTER_PT (Left)
        p = _get_effective_path("FOOTER_PT", ["PMI갑", "PMI-1", "PT"])
        if p: self.place_image_freely(ws, p, _get_val("FOOTER_PT", "_ANCHOR", "A1"), float(_get_val("FOOTER_PT", "_W", 100)), float(_get_val("FOOTER_PT", "_H", 50)), float(_get_val("FOOTER_PT", "_X", 0)), float(_get_val("FOOTER_PT", "_Y", 0)))



    def force_print_settings(self, ws, context="DATA"):
        try:
            # [ENHANCED] Use tracked current_mode
            mode = getattr(self, 'current_mode', "PMI")
            try:
                tab_text = self.mode_notebook.tab(self.mode_notebook.select(), "text")
                if "PMI" in tab_text: mode = "PMI"
                elif "RT" in tab_text: mode = "RT"
                elif "PT" in tab_text: mode = "PT"
                elif "PAUT" in tab_text: mode = "PAUT"
            except: pass
            
            full_context = f"{mode}_{context}"

            # [NEW] Manual Print Area Override (Priority: Mode-Specific > Global)
            manual_area = self.config.get(f'PRINT_AREA_{full_context}', "").strip()
            if not manual_area:
                manual_area = self.config.get(f'PRINT_AREA_{context}', "").strip()
                
            if manual_area:
                ws.print_area = manual_area
            else:
                if context == "COVER":
                    # [NEW] Highly Dynamic print area for Gapji
                    key = f"{mode}_GAPJI_PRINT_END_ROW" if mode != "PMI" else "GAPJI_PRINT_END_ROW"
                    end_r = int(self.config.get(key, 51))
                    
                    # [FIX] 사용자의 요청에 따라 RT 갑지의 인쇄 영역을 A1:P49로 강제 최적화
                    if mode == "RT":
                        ws.print_area = 'A1:P49'
                    elif end_r > 0:
                        ws.print_area = f'A1:T{end_r}'
                else:
                    # [NEW] Highly Dynamic print area for Eulji
                    key = f"{mode}_PRINT_END_ROW" if mode != "PMI" else "PRINT_END_ROW"
                    end_r = int(self.config.get(key, 47))
                    if end_r > 0:
                        ws.print_area = f'A1:M{end_r}'
            
            # [FIX] Clear any manual page breaks that might cause mid-page splitting
            try:
                ws.row_breaks = []
                ws.col_breaks = []
            except: pass

            ws.page_setup.paperSize = 9; ws.page_setup.orientation = 'portrait'
            
            # [FIX] 모드별 설정 우선 읽기 (RT_COVER → COVER 순 폴백)
            default_scale = 95
            if mode == "RT" and context == "COVER": default_scale = 90
            def _cfg_str(key):
                v = self.config.get(key, '')
                return str(v).strip() if v != '' else ''
            scale_val = (_cfg_str(f'PRINT_SCALE_{full_context}') or
                         _cfg_str(f'PRINT_SCALE_{context}') or
                         str(default_scale))
            ws.page_setup.scale = int(float(scale_val))
            ws.print_options.horizontalCentered = True; ws.print_options.verticalCentered = True
            
            def _margin(name, default):
                return float(
                    _cfg_str(f'MARGIN_{full_context}_{name}') or
                    _cfg_str(f'MARGIN_{context}_{name}') or
                    default
                )
            ws.page_margins.top    = _margin('TOP',    0.2)
            ws.page_margins.bottom = _margin('BOTTOM', 0.2)
            ws.page_margins.left   = _margin('LEFT',   0.5)
            ws.page_margins.right  = _margin('RIGHT',  0.3)

        except Exception as e:
            print(f"[WARNING] Print settings failed: {e}")
            print(f"[WARNING] Print settings failed: {e}")

    def apply_custom_dimensions(self, ws, context):
        """[NEW] 사용자의 행/열 조절 설정을 파싱하여 적용 (모드별 설정 대응)"""
        try:
            # Determine current mode for key fallback
            mode = getattr(self, 'current_mode', "PMI")
            try:
                tab_text = self.mode_notebook.tab(self.mode_notebook.select(), "text")
                if "PMI" in tab_text: mode = "PMI"
                elif "RT" in tab_text: mode = "RT"
                elif "PT" in tab_text: mode = "PT"
                elif "PAUT" in tab_text: mode = "PAUT"
            except: pass
            
            full_context = f"{mode}_{context}" if "_" not in context else context

            # Row Adjustment
            row_range_str = (self.config.get(f"CUSTOM_ROWS_{full_context}", "") or 
                             self.config.get(f"CUSTOM_ROWS_{context}", "")).strip()
            
            if row_range_str:
                height_val = (self.config.get(f"CUSTOM_ROW_HEIGHT_{full_context}") or 
                              self.config.get(f"CUSTOM_ROW_HEIGHT_{context}", 16.5))
                height = float(height_val)
                for part in row_range_str.split(','):
                    part = part.strip()
                    if not part: continue
                    if '-' in part or '~' in part:
                        sep = '-' if '-' in part else '~'
                        start, end = map(int, part.split(sep))
                        for r in range(start, end + 1):
                            ws.row_dimensions[r].height = height
                    else:
                        ws.row_dimensions[int(part)].height = height
            
            # Column Adjustment
            col_range_str = (self.config.get(f"CUSTOM_COLS_{full_context}", "") or 
                             self.config.get(f"CUSTOM_COLS_{context}", "")).strip()
            if col_range_str:
                width_val = (self.config.get(f"CUSTOM_COL_WIDTH_{full_context}") or 
                             self.config.get(f"CUSTOM_COL_WIDTH_{context}", 10.0))
                width = float(width_val)
                for part in col_range_str.split(','):
                    if '-' in part:
                        bits = part.split('-')
                        if len(bits) == 2:
                            start_l, end_l = bits
                            for c in range(column_index_from_string(start_l), column_index_from_string(end_l) + 1):
                                ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = width
                    else:
                        ws.column_dimensions[part.strip().upper()].width = width
        except Exception as e:
            print(f"[DEBUG] Custom dimension failed for {context}: {e}")

    def safe_set_value(self, ws, coord, value, align=None):
        """병합된 셀이라도 기준 셀(Top-Left)을 찾아 안전하게 값을 입력합니다. (성능 최적화 버전)"""
        try:
            # 전달받은 인자가 이미 셀 객체인 경우
            if hasattr(coord, 'coordinate'):
                target_cell = coord
                coord_str = target_cell.coordinate
            else:
                target_cell = ws[coord]
                coord_str = coord

            # 해당 셀이 MergedCell(병합 영역의 기준 셀이 아닌 곳)인 경우에만 탐색 수행
            if isinstance(target_cell, MergedCell):
                for m_range in ws.merged_cells.ranges:
                    if coord_str in m_range:
                        target_cell = ws.cell(row=m_range.min_row, column=m_range.min_col)
                        break
            
            target_cell.value = value
            if align:
                target_cell.alignment = Alignment(horizontal=align, vertical='center')
        except Exception: pass

    def safe_merge_cells(self, ws, start_row, start_column, end_row, end_column):
        """[FIX] Prevent unmerging or conflicting merges on the cover sheet"""
        try:
            # If it's the cover sheet, be extremely conservative with merging/unmerging
            if ws == ws.parent.worksheets[0]:
                return
        except: pass
        
        try:
            ws.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
        except Exception as e:
            self.log(f"   ⚠️ 셀 병합 실패 ({start_row},{start_column}): {e}")

    def set_eulji_headers(self, ws):
        # [FIX] RT 모드일 때는 헤더를 자동으로 쓰지 않음 (템플릿 보존)
        if getattr(self, 'current_mode', "") == "RT":
            return

        headers = ["NI", "CR", "MO"]
        data_font = Font(size=9); header_row = self.config.get('START_ROW', 10)
        for i, val in enumerate(headers):
            col = 8 + i
            cell = ws.cell(row=header_row, column=col)
            self.safe_set_value(ws, cell.coordinate, val, align='center')
            cell.font = data_font
        
        materials = "SS304,SS304L,SS316,SS316L,SS321,SS347,SS410,SS430,DUPLEX,MONEL,INCONEL,ER308,ER308L,ER309,ER309L,ER316,ER316L,ER347,ER2209,WP316,WP316L,TP316,TP316L,F316L,A182-F316L,A312-TP316L"
        dv_q = DataValidation(type="list", formula1=f'"{materials}"', allow_blank=True)
        ws.add_data_validation(dv_q)
        for r in range(self.config['START_ROW'], self.config['DATA_END_ROW'] + 1):
            target_l = ws.cell(row=r, column=13); dv_q.add(target_l) # 12 -> 13 (M)
            target_l.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center'); target_l.font = Font(size=8.5)
    
        # [NEW] K15에 하이픈 추가
        self.safe_set_value(ws, 'K15', "-")
        ws['K15'].alignment = Alignment(horizontal='center', vertical='center')
        
    def prepare_next_sheet(self, wb, source_sheet_idx, page_num):
        source_sheet = wb.worksheets[source_sheet_idx]; new_sheet = wb.copy_worksheet(source_sheet) 
        
        # [FIX] openpyxl's copy_worksheet does not copy images. Manually copy them for RT to preserve Shooting Sketches.
        if getattr(self, 'current_mode', "") == "RT":
            try:
                import copy
                for img in source_sheet._images:
                    new_img = copy.copy(img)
                    if hasattr(img, 'anchor'):
                        new_img.anchor = copy.copy(img.anchor)
                    new_sheet.add_image(new_img)
            except: pass
            
        base_title = source_sheet.title.split('_')[0]; new_sheet.title = f"{base_title[:20]}_{page_num:03d}"
        self.force_print_settings(new_sheet, context="DATA")
        
        # [FIX] Only add logos if they weren't already copied from the source sheet
        if getattr(self, 'current_mode', "") != "RT" or not new_sheet._images:
            self.add_logos_to_sheet(new_sheet, is_cover=False)
        self.apply_custom_dimensions(new_sheet, "DATA")
        for col_letter, col_dim in source_sheet.column_dimensions.items(): new_sheet.column_dimensions[col_letter].width = col_dim.width
        data_font = Font(size=9); grade_font = Font(size=8.5)
        for r in range(self.config['START_ROW'], self.config['DATA_END_ROW'] + 1):
            rd = new_sheet.row_dimensions[r] # [REMOVED] Hardcoded 20.55 override
        
        # [FIX] RT 모드일 경우 Shooting Sketch(32행~42행)를 보호하기 위해 데이터 종료 행을 조절
        current_mode = getattr(self, 'current_mode', "")
        start_row = self.config.get('START_ROW', 17)
        end_row = self.config.get('DATA_END_ROW', 45)
        
        if current_mode == "RT":
            start_row = int(self.config.get('RT_START_ROW', 17))
            end_row = int(self.config.get('RT_END_ROW', 31))
            
        for r in range(start_row + 1, end_row + 1):
            for c in range(1, 14):
                cell = new_sheet.cell(row=r, column=c)
                cell.font = grade_font if c == 13 else data_font
                self.safe_set_value(new_sheet, cell, None)
                
        merged_to_clear = [rng for rng in new_sheet.merged_cells.ranges if rng.min_row >= start_row and rng.max_row <= end_row]
        for rng in merged_to_clear: new_sheet.unmerge_cells(str(rng))
        
        # [FIX] 갑지 데이터 수식으로 연결 (첫번째 시트 참조)
        try:
            ws0 = wb.worksheets[0]
            if len(wb.worksheets) > 1:
                # 엑셀 수식을 사용하여 갑지의 값이 바뀌면 자동으로 바뀌게 설정
                self.safe_set_value(new_sheet, 'K5', f"='{ws0.title}'!L5")
                self.safe_set_value(new_sheet, 'M5', f"='{ws0.title}'!N5")
                self.safe_set_value(new_sheet, 'M8', f"='{ws0.title}'!N8")
                # [FIX] K5:M10 범위 글씨체 바탕, 크기 9 적용
                for r_idx in range(5, 11):
                    for c_idx in range(11, 14): # K(11) ~ M(13)
                        cell = new_sheet.cell(row=r_idx, column=c_idx)
                        cell.font = Font(name='바탕', size=9, bold=False)
                        # K5, K8는 줄바꿈 적용, 그 외는 셀 크기에 맞춤 적용
                        if (r_idx == 5 or r_idx == 8) and c_idx == 11:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        except: pass
        
        # [NEW] K15에 하이픈 추가
        self.safe_set_value(new_sheet, 'K15', "-")
        new_sheet['K15'].alignment = Alignment(horizontal='center', vertical='center')
        
        new_sheet.column_dimensions['M'].width = 18.0
        return new_sheet

    def inject_drawing_layer(self, template_path, target_path):
        """
        [FINAL V4] Definitive hybrid surgery.
        - Report's workbook.xml, workbook.xml.rels, [Content_Types].xml → correct structure, print area, all pages
        - Template's drawings, media, sheet structure → drawings preserved
        - Report's sheetData injected into template's sheet XMLs → content preserved
        - Blueprint for extra sheets (multipage RT)
        """
        import zipfile, shutil, os, re, traceback

        final_out = target_path + ".final"
        shutil.copy2(template_path, final_out)
        temp_swap = final_out + ".swap"

        try:
            # ── Step 1: 리포트에서 필요한 모든 데이터 수집 ──────────────────
            with zipfile.ZipFile(target_path, 'r') as z_rep:
                rep_names = z_rep.namelist()

                # sharedStrings 읽기
                rep_shared_strings = []
                ss_f = next((f for f in rep_names if f.lower() == 'xl/sharedstrings.xml'), None)
                if ss_f:
                    try:
                        ss_xml = z_rep.read(ss_f).decode('utf-8', errors='ignore')
                        rep_shared_strings = re.findall(r'<t(?:\s[^>]*)?>([^<]*)</t>', ss_xml)
                    except: pass

                # 시트 데이터 추출 (sharedStr → inlineStr 변환)
                processed_sheets = {}
                for f in rep_names:
                    if f.lower().startswith('xl/worksheets/sheet') and f.lower().endswith('.xml'):
                        try:
                            content = z_rep.read(f).decode('utf-8', errors='ignore')
                            def _to_inline(m, _ss=rep_shared_strings):
                                try:
                                    idx = int(m.group(2))
                                    if 0 <= idx < len(_ss):
                                        t = _ss[idx].replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
                                        return f'<c {m.group(1).replace(chr(116)+"=\"s\"","t=\"inlineStr\"")}><is><t>{t}</t></is>'
                                except: pass
                                return m.group(0)
                            content = re.sub(r'<c ([^>]*t="s"[^>]*)>\s*<v>(\d+)</v>', _to_inline, content)
                            processed_sheets[f.lower()] = content
                        except: pass

                # 리포트 pageSetup & pageMargins 추출 (force_print_settings 적용값)
                rep_page_settings = {}  # fl_lower → {'setup': '...', 'margins': '...'}
                rep_scale = None
                for fl_s, c in processed_sheets.items():
                    setup_m  = re.search(r'<pageSetup\b[^>]*/>', c)
                    margin_m = re.search(r'<pageMargins\b[^>]*/>', c)
                    scale_m  = re.search(r'<pageSetup\b[^>]*\bscale="(\d+)"', c)
                    if scale_m and not rep_scale:
                        rep_scale = scale_m.group(1)
                    rep_page_settings[fl_s] = {
                        'setup':   setup_m.group(0)  if setup_m  else None,
                        'margins': margin_m.group(0) if margin_m else None,
                    }

                # 리포트의 workbook 메타 파일들 (raw bytes)
                rep_workbook_xml  = next((z_rep.read(f) for f in rep_names if f.lower() == 'xl/workbook.xml'), None)
                rep_wb_rels       = next((z_rep.read(f) for f in rep_names if f.lower() == 'xl/_rels/workbook.xml.rels'), None)
                rep_content_types = next((z_rep.read(f) for f in rep_names if f.lower() == '[content_types].xml'), None)
                rep_styles        = next((z_rep.read(f) for f in rep_names if f.lower() == 'xl/styles.xml'), None)
                rep_ss_data       = next((z_rep.read(f) for f in rep_names if f.lower() == 'xl/sharedstrings.xml'), None)

            # ── Step 2: 템플릿에서 드로잉/Blueprint 수집 ────────────────────
            with zipfile.ZipFile(final_out, 'r') as z_tmpl:
                blueprint_xml  = None
                blueprint_rels = None
                first_sheet_f  = next((f for f in z_tmpl.namelist()
                                       if f.lower().startswith('xl/worksheets/sheet') and f.lower().endswith('.xml')), None)
                if first_sheet_f:
                    blueprint_xml = z_tmpl.read(first_sheet_f).decode('utf-8', errors='ignore')
                    rel_p = f"xl/worksheets/_rels/{os.path.basename(first_sheet_f)}.rels"
                    if rel_p in z_tmpl.namelist():
                        blueprint_rels = z_tmpl.read(rel_p)

                # 템플릿 파일 맵 (소문자 경로 → 파일 내용)
                tmpl_map = {}
                for item in z_tmpl.infolist():
                    tmpl_map[item.filename.lower()] = z_tmpl.read(item.filename)

            # ── Step 3: 출력 ZIP 구성 ────────────────────────────────────────
            with zipfile.ZipFile(temp_swap, 'w', compression=zipfile.ZIP_DEFLATED) as z_out:

                # ── 인쇄영역 강제 확정 주입 (시트별 구분: 갑지 vs 을지)
                if rep_workbook_xml:
                    wb_str = rep_workbook_xml.decode('utf-8', errors='ignore')
                    snames = re.findall(r'<sheet\b[^>]*\bname="([^"]+)"', wb_str)
                    
                    # 기존 Print_Area definedName 전체 제거
                    wb_str = re.sub(r'<definedName[^>]*name="_xlnm\.Print_Area"[^/]*/>', '', wb_str)
                    wb_str = re.sub(r'<definedName[^>]*name="_xlnm\.Print_Area"[^>]*>.*?</definedName>\s*',
                                    '', wb_str, flags=re.DOTALL)
                    
                    def _get_norm_pa(pa_str):
                        return re.sub(r'\$?([A-Za-z]+)\$?(\d+)', lambda m2: f'${m2.group(1).upper()}${m2.group(2)}', pa_str)

                    new_dns_list = []
                    for i, s_name in enumerate(snames):
                        if i == 0: # 갑지 (Cover)
                            pa = (self.config.get('PRINT_AREA_RT_COVER', '').strip() or 
                                  self.config.get('PRINT_AREA_COVER', '').strip() or 'A1:P49')
                        else: # 을지 (Data)
                            pa = (self.config.get('PRINT_AREA_RT_DATA', '').strip() or 
                                  self.config.get('PRINT_AREA_DATA', '').strip() or 'A1:M47')
                        
                        pa_norm = _get_norm_pa(pa)
                        new_dns_list.append(f'<definedName name="_xlnm.Print_Area" localSheetId="{i}">\'{s_name}\'!{pa_norm}</definedName>')
                    
                    new_dns = ''.join(new_dns_list)
                    if '<definedNames' in wb_str:
                        wb_str = re.sub(r'(<definedNames[^>]*>)', r'\1' + new_dns, wb_str, count=1)
                    elif '</workbook>' in wb_str:
                        wb_str = wb_str.replace('</workbook>', f'<definedNames>{new_dns}</definedNames></workbook>', 1)
                    
                    rep_workbook_xml = wb_str.encode('utf-8')
                    self.log(f"   📐 [인쇄영역] 갑지/을지 개별 영역 주입 완료")
                
                # [Content_Types].xml: 템플릿 기반 + 리포트의 추가 시트 Override 병합
                # (템플릿 버전: drawing, media 등 포함 / 리포트 버전: 추가 시트 항목 포함)
                merged_ct = tmpl_map.get('[content_types].xml', b'')
                if merged_ct and rep_content_types:
                    tmpl_ct_str = merged_ct.decode('utf-8', errors='ignore')
                    rep_ct_str  = rep_content_types.decode('utf-8', errors='ignore')
                    # 템플릿에 없는 Override 항목을 리포트에서 추가
                    tmpl_parts  = set(re.findall(r'PartName="([^"]+)"', tmpl_ct_str))
                    for m_ov in re.finditer(r'<Override\b[^/]*/>', rep_ct_str):
                        pn_m = re.search(r'PartName="([^"]+)"', m_ov.group(0))
                        if pn_m and pn_m.group(1) not in tmpl_parts:
                            tmpl_ct_str = tmpl_ct_str.replace('</Types>', m_ov.group(0) + '</Types>', 1)
                    merged_ct = tmpl_ct_str.encode('utf-8')
                z_out.writestr('[Content_Types].xml', merged_ct if merged_ct else (rep_content_types or b''))

                # 리포트의 workbook 메타 파일 기록
                if rep_workbook_xml:  z_out.writestr('xl/workbook.xml',           rep_workbook_xml)
                if rep_wb_rels:       z_out.writestr('xl/_rels/workbook.xml.rels', rep_wb_rels)
                if rep_styles:        z_out.writestr('xl/styles.xml',              rep_styles)
                if rep_ss_data:       z_out.writestr('xl/sharedStrings.xml',       rep_ss_data)
                self.log("   📋 [workbook] 리포트 workbook 구조 적용")


                written = {'xl/workbook.xml', 'xl/_rels/workbook.xml.rels',
                           '[content_types].xml', 'xl/styles.xml', 'xl/sharedstrings.xml'}

                # 템플릿의 모든 파일 처리 (sheet, drawing, media 등)
                for fl_lower, raw in tmpl_map.items():
                    if fl_lower in written:
                        continue  # 이미 리포트 버전으로 기록됨

                    if fl_lower.startswith('xl/worksheets/sheet') and fl_lower.endswith('.xml'):
                        # 시트 XML: 템플릿 구조 + 리포트 데이터
                        try:
                            tmpl_sheet = raw.decode('utf-8', errors='ignore')
                            # rowBreaks 제거
                            tmpl_sheet = re.sub(r'<rowBreaks[^>]*>.*?</rowBreaks>', '', tmpl_sheet, flags=re.DOTALL)
                            tmpl_sheet = re.sub(r'<rowBreaks[^>]*/>', '', tmpl_sheet)
                            tmpl_sheet = re.sub(r'<colBreaks[^>]*>.*?</colBreaks>', '', tmpl_sheet, flags=re.DOTALL)
                            # pageSetup 전체 교체 (배율 + 방향 + 용지 크기)
                            rep_ps = rep_page_settings.get(fl_lower, {}).get('setup')
                            if rep_ps:
                                if re.search(r'<pageSetup\b[^>]*/>', tmpl_sheet):
                                    tmpl_sheet = re.sub(r'<pageSetup\b[^>]*/>', rep_ps, tmpl_sheet, count=1)
                                else:
                                    # pageSetup 태그가 없으면 sheetData 앞에 삽입
                                    tmpl_sheet = tmpl_sheet.replace('</sheetData>', '</sheetData>' + rep_ps, 1)
                            elif rep_scale:
                                # pageSetup은 없지만 scale만 있으면 기존 방식으로 패치
                                tmpl_sheet = re.sub(r'<pageSetup\b[^>]*/>', lambda m: re.sub(r'\bscale="\d+"', f'scale="{rep_scale}"', m.group(0)) if 'scale=' in m.group(0) else m.group(0).replace('<pageSetup', f'<pageSetup scale="{rep_scale}"', 1), tmpl_sheet)
                            
                            # pageMargins 교체 (여백 설정)
                            rep_pm = rep_page_settings.get(fl_lower, {}).get('margins')
                            if rep_pm:
                                if re.search(r'<pageMargins\b[^>]*/>', tmpl_sheet):
                                    tmpl_sheet = re.sub(r'<pageMargins\b[^>]*/>', rep_pm, tmpl_sheet, count=1)
                                else:
                                    tmpl_sheet = tmpl_sheet.replace('</pageSetup>', '</pageSetup>' + rep_pm)
                            # sheetData 주입 + 행 높이 보존
                            if fl_lower in processed_sheets:
                                rc = processed_sheets[fl_lower]
                                s, e = '<sheetData>', '</sheetData>'
                                si, ei = rc.find(s), rc.find(e)
                                ts, te = tmpl_sheet.find(s), tmpl_sheet.find(e)
                                if si != -1 and ei != -1 and ts != -1 and te != -1:
                                    # 1) 템플릿의 행 높이/속성 추출 (ht, customHeight, hidden)
                                    tmpl_sd = tmpl_sheet[ts+len(s):te]
                                    tmpl_row_attrs = {}
                                    for rm in re.finditer(r'<row\b([^>]*)>', tmpl_sd):
                                        ra = rm.group(1)
                                        rid = re.search(r'\br="(\d+)"', ra)
                                        if rid:
                                            tmpl_row_attrs[rid.group(1)] = ra

                                    # 2) 리포트 sheetData 주입
                                    tmpl_sheet = tmpl_sheet[:ts+len(s)] + rc[si+len(s):ei] + tmpl_sheet[te:]

                                    # 3) 템플릿 행 높이 복원 (ht, customHeight 속성만 덮어씀)
                                    def _restore_height(m, _ra=tmpl_row_attrs):
                                        attrs = m.group(1)
                                        rid = re.search(r'\br="(\d+)"', attrs)
                                        if rid and rid.group(1) in _ra:
                                            ta = _ra[rid.group(1)]
                                            for attr in ('ht', 'customHeight', 'hidden'):
                                                # [FIX] If ht or customHeight already exists in attrs (user set it), skip overwriting from template
                                                if attr in ('ht', 'customHeight') and f'{attr}=' in attrs:
                                                    continue
                                                
                                                am = re.search(fr'\b{attr}="([^"]+)"', ta)
                                                if am:
                                                    val = f'{attr}="{am.group(1)}"'
                                                    if f'{attr}=' in attrs:
                                                        attrs = re.sub(fr'\b{attr}="[^"]*"', val, attrs)
                                                    else:
                                                        attrs = attrs.rstrip() + f' {val}'
                                        return f'<row {attrs}>'
                                    # 주입된 sheetData 영역에만 적용
                                    ts2 = tmpl_sheet.find(s)
                                    te2 = tmpl_sheet.find(e)
                                    if ts2 != -1 and te2 != -1:
                                        sd_inner = tmpl_sheet[ts2+len(s):te2]
                                        sd_inner = re.sub(r'<row\b([^>]*)>', _restore_height, sd_inner)
                                        tmpl_sheet = tmpl_sheet[:ts2+len(s)] + sd_inner + tmpl_sheet[te2:]
                                    self.log(f"   ✅ [V4] {os.path.basename(fl_lower)} 데이터 주입 (행높이 보존)")
                            z_out.writestr(fl_lower, tmpl_sheet.encode('utf-8'))

                        except:
                            z_out.writestr(fl_lower, raw)
                    else:
                        # 나머지 (drawing, media, rels, etc.) → 템플릿 그대로
                        z_out.writestr(fl_lower, raw)
                    written.add(fl_lower)

                # 추가 시트 (템플릿에 없는 멀티페이지 시트) → Blueprint 적용
                for fl_rep, rc in processed_sheets.items():
                    if fl_rep in written:
                        continue
                    if blueprint_xml:
                        try:
                            s, e = '<sheetData>', '</sheetData>'
                            si, ei = rc.find(s), rc.find(e)
                            bp = re.sub(r'<rowBreaks[^>]*>.*?</rowBreaks>', '', blueprint_xml, flags=re.DOTALL)
                            ts, te = bp.find(s), bp.find(e)
                            if si != -1 and ei != -1 and ts != -1 and te != -1:
                                new_xml = bp[:ts+len(s)] + rc[si+len(s):ei] + bp[te:]
                                z_out.writestr(fl_rep, new_xml.encode('utf-8'))
                                if blueprint_rels:
                                    z_out.writestr(f"xl/worksheets/_rels/{os.path.basename(fl_rep)}.rels", blueprint_rels)
                                self.log(f"   📄 [V4] {os.path.basename(fl_rep)} Blueprint 추가")
                                written.add(fl_rep)
                        except: pass

            os.remove(target_path)
            os.rename(temp_swap, target_path)
            if os.path.exists(final_out): os.remove(final_out)
            self.log("   🛡️ [V4 완료] 인쇄영역 + 그림 + 내용 모두 보존")

        except Exception as e:
            for fc in [temp_swap, final_out]:
                if os.path.exists(fc):
                    try: os.remove(fc)
                    except: pass
            self.log(f"   ⚠️ [오류] 드로잉 수술 실패: {e}")
            traceback.print_exc()

    def calculate_rt_shots(self, size_str):
        """배관 구경을 분석하여 [표 1] 기준 촬영 매수 반환"""
        try:
            nums = re.findall(r"[-+]?\d*\.\d+|\d+", str(size_str))
            if not nums: return 1
            val = float(nums[0])
            if val >= 30: return 7
            if val >= 26: return 6
            if val >= 20: return 5
            if val == 18: return 6
            if val >= 14: return 5
            if val >= 10: return 4
            if val >= 6: return 3
            if val >= 3: return 3
            if val <= 2.5: return 2
            return 1
        except: return 1

    def extract_only(self, show_msg=True):
        """데이터만 추출하여 리스트와 미리보기에 반영 (PMI/RT 대응)"""
        self.save_settings() # [FIX] UI 설정(컬럼명 등)을 즉시 반영하여 리셋 방지
        # 현재 활성 탭에 따른 모드 결정
        try:
            main_tab = self.mode_notebook.tab(self.mode_notebook.select(), "text")
            if "RT" in main_tab:
                # [FIX] 하부 탭 이름을 직접 확인하여 모드 결정
                sub_tab_text = self.rt_preview_nb.tab(self.rt_preview_nb.select(), "text")
                mode = "KOGAS" if "가스공사" in sub_tab_text else "RT"
            elif "PT" in main_tab: mode = "PT"
            elif "PAUT" in main_tab: mode = "PAUT"
            else: mode = "PMI"
        except Exception as e:
            self.log(f"⚠️ 모드 판별 오류: {e}")
            mode = "PMI"

        if mode == "PT":
            target_file = self.pt_target_file_path.get()
        elif mode == "PAUT":
            target_file = self.paut_target_file_path.get()
        elif mode == "RT":
            target_file = self.rt_target_file_path.get()
        elif mode == "KOGAS":
            target_file = self.kogas_target_file_path.get()
        else:
            target_file = self.target_file_path.get()
        if not target_file:
            messagebox.showwarning("파일 미선택", f"{mode} 데이터 파일을 선택해주세요.")
            return False
            
        self.log(f"🔍 {mode} 데이터 추출 시작: {os.path.basename(target_file)}")
        
        # [NEW] Extract date from filename
        fname = os.path.basename(target_file)
        date_match = re.search(r'(\d{4}[-._]\d{2}[-._]\d{2}|\d{2}[-._]\d{2}[-._]\d{2}|\d{8}|\d{6})', fname)
        extracted_date = date_match.group(0) if date_match else ""
        
        # Standardize formatting
        if extracted_date:
            clean_date = re.sub(r'[-._]', '', extracted_date)
            if len(clean_date) == 8:
                extracted_date = f"{clean_date[:4]}-{clean_date[4:6]}-{clean_date[6:]}"
            elif len(clean_date) == 6:
                extracted_date = f"20{clean_date[:2]}-{clean_date[2:4]}-{clean_date[4:]}"
        
        if not extracted_date:
            extracted_date = datetime.datetime.now().strftime("%Y-%m-%d")
            self.log(f"⚠️ 파일명에서 날짜를 찾지 못해 오늘 날짜({extracted_date})로 설정합니다.")
        else:
            self.log(f"📅 파일 날짜 인식 성공: {extracted_date}")

        self.progress['value'] = 0
        all_extracted_data = []

        # [AUTO COLUMN ADDITION] 
        if mode == "PMI" and hasattr(self, 'element_filters') and self.element_filters:
            new_keys = []
            for f_item in self.element_filters:
                f_k = f_item['key'].get().strip()
                if f_k and f_k.upper() not in ["CR", "NI", "MO", "MN"] and f_k.capitalize() not in self.column_keys and f_k.upper() not in self.column_keys and f_k not in self.column_keys:
                    new_keys.append(f_k)
            
            if new_keys:
                self.column_keys.extend(new_keys)
                current_cols = list(self.preview_tree["columns"])
                current_display = list(self.preview_tree.cget("displaycolumns"))
                if current_display[0] == "#all": current_display = list(current_cols)

                # Save existing configurations to survive the tree["columns"] reset
                configs = {}
                for c in current_cols:
                    configs[c] = {
                        "text": self.preview_tree.heading(c).get("text", c),
                        "width": self.preview_tree.column(c).get("width", 80)
                    }

                for k in new_keys:
                    if k not in current_cols: current_cols.append(k)
                    if k not in current_display: current_display.append(k)
                
                self.preview_tree["columns"] = tuple(current_cols)
                self.preview_tree["displaycolumns"] = tuple(current_display)
                
                # Re-apply all configs + new keys setup
                for c in current_cols:
                    header_txt = configs.get(c, {}).get("text", c)
                    w = configs.get(c, {}).get("width", 80)
                    self.preview_tree.heading(c, text=header_txt, anchor='center', command=lambda _c=c: self.sort_by_column(_c, mode="PMI"))
                    self.preview_tree.column(c, width=w, anchor='center', stretch=False)
        
        def _find_col(df, keywords, exclude=None):
            for col in df.columns:
                c_up = str(col).upper().strip()
                if exclude and any(ex in c_up for ex in exclude): continue
                if any(k == c_up for k in keywords): return col
            for col in df.columns:
                c_up = str(col).upper().strip()
                if exclude and any(ex in c_up for ex in exclude): continue
                if any(k in c_up for k in keywords):
                    if "NI" in keywords and ("UNIT" in c_up or "LINE" in c_up): continue
                    return col
            return None

        def _get_kw(k): 
            if hasattr(self, 'rt_extract_mappings') and k in self.rt_extract_mappings:
                return [x.strip().upper() for x in self.rt_extract_mappings[k].get().split(',')]
            return []

        try:
            target_input = self.sequence_filter.get().strip()
            target_no_list = [x.strip() for x in target_input.replace(',', ' ').split() if target_input and x.strip()] if target_input else []
            
            with pd.ExcelFile(target_file) as xls:
                for s_idx, sheet_name in enumerate(xls.sheet_names):
                    self.log(f"📄 시트 스캔: {sheet_name}")
                    try: temp_df = pd.read_excel(xls, sheet_name=sheet_name, header=None, nrows=50)
                    except: continue
                    
                    # [지능형 시트 내 날짜 추출]
                    # 의뢰서의 L5, L6 셀(머지된 영역)을 직접 파싱하여 
                    # 파일명 날짜보다 우선하여 해당 시트의 기본 날짜(sheet_level_date)로 활용합니다.
                    sheet_level_date = extracted_date
                    try:
                        for r_pos in [4, 5]: # 0-indexed 4=5행, 5=6행
                            for c_pos in [11, 10, 12]: # 11=L열, 10=K열, 12=M열
                                if r_pos < len(temp_df) and c_pos < len(temp_df.columns):
                                    val = temp_df.iloc[r_pos, c_pos]
                                    if pd.notna(val):
                                        if isinstance(val, (datetime.datetime, datetime.date, pd.Timestamp)) or hasattr(val, 'strftime'):
                                            sheet_level_date = val.strftime('%Y-%m-%d')
                                            break
                                        cleaned = str(val).strip()
                                        # YYYY-MM-DD 혹은 YYYY.MM.DD 등의 형태에서 날짜만 클리닝
                                        cleaned_num = re.sub(r'[^0-9]', '', cleaned)
                                        if len(cleaned_num) == 8:
                                            sheet_level_date = f"{cleaned_num[:4]}-{cleaned_num[4:6]}-{cleaned_num[6:]}"
                                            break
                                        elif len(cleaned_num) == 6:
                                            prefix = "20" if int(cleaned_num[:2]) < 50 else "19"
                                            sheet_level_date = f"{prefix}{cleaned_num[:2]}-{cleaned_num[2:4]}-{cleaned_num[4:]}"
                                            break
                            if sheet_level_date != extracted_date: break
                    except: pass
                    
                    if sheet_level_date != extracted_date:
                        self.log(f"   📅 의뢰서 헤더(L5/L6 셀)에서 날짜 파싱 성공: {sheet_level_date}")
                        
                    # [지능형 시트 내 도면번호 추출]
                    # 의뢰서의 C5, C6 셀(머지된 영역)을 직접 파싱하여
                    # 해당 시트의 기본 도면번호(sheet_level_dwg)로 활용합니다.
                    sheet_level_dwg = ""
                    try:
                        for r_pos in [4, 5]: # 0-indexed 4=5행, 5=6행
                            for c_pos in [2, 1, 3]: # 2=C열, 1=B열, 3=D열
                                if r_pos < len(temp_df) and c_pos < len(temp_df.columns):
                                    val = temp_df.iloc[r_pos, c_pos]
                                    if pd.notna(val):
                                        val_str = str(val).strip()
                                        if val_str and val_str.lower() != 'nan' and len(val_str) > 2:
                                            # "도면번호", "도면", "DWG" 같은 라벨 텍스트는 제외
                                            if val_str in ["도면번호", "도면", "DWG", "DWG NO", "DWG.NO", "도면 NO", "도면 NO."]:
                                                continue
                                            # "SEOUL INSPECTION...", 회사명 혹은 주소 등 공백이 너무 많은 텍스트 제외
                                            val_upper = val_str.upper()
                                            if any(k in val_upper for k in ["CO.", "LTD", "INSPECTION", "TESTING", "CORP", "INC", "SEOUL", "주식회사"]):
                                                continue
                                            if val_str.count(" ") > 2:
                                                continue
                                            sheet_level_dwg = val_str
                                            break
                            if sheet_level_dwg: break
                    except: pass
                    
                    if sheet_level_dwg:
                        self.log(f"   📐 의뢰서 헤더(C5/C6 셀)에서 도면번호 파싱 성공: {sheet_level_dwg}")
                    
                    header_idx = None
                    is_fixed_range = False
                    # [NEW] KOGAS 고정 영역 처리 (A9:AB13 헤더, A14:AB25 데이터)
                    if mode == "KOGAS":
                        header_idx = 8 # Excel Row 9
                        is_fixed_range = True
                    else:
                        for i, row in temp_df.iterrows():
                            row_str = str(row.values).upper()
                            if mode == "RT":
                                if any(k in row_str for k in ["FILM", "DEFECT", "IQI", "순번", "용접부"]):
                                    header_idx = i; break
                            elif mode == "PT":
                                if (("LINE" in row_str or "ISO" in row_str or "DWG" in row_str) and ("JOINT" in row_str or "WELD" in row_str)):
                                    header_idx = i; break
                            else:
                                if ("CR" in row_str and "NI" in row_str) or ("CHROMIUM" in row_str):
                                    header_idx = i; break
                    
                    if header_idx is None: continue
                    
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=header_idx)
                    if df.empty: continue

                    if is_fixed_range:
                        # [KOGAS] 9~13행 헤더 병합 (df.columns + df.iloc[0:3])
                        new_cols = []
                        for col_idx in range(len(df.columns)):
                            parts = [str(df.columns[col_idx])]
                            for r_idx in range(0, 4): # 10, 11, 12, 13행
                                val = df.iloc[r_idx, col_idx]
                                if pd.notna(val) and "UNNAMED" not in str(val).upper():
                                    parts.append(str(val))
                            new_cols.append(" ".join(parts).upper())
                        df.columns = new_cols
                        # [KOGAS] 14~25행 데이터 추출 (인덱스 4~15)
                        df = df.iloc[4:16].reset_index(drop=True)
                    else:
                        # [STANDARD] 기존 자동 보정 로직 (두 줄 헤더 등)
                        row0_str = " ".join([str(x).upper() for x in df.iloc[0] if pd.notna(x)])
                        if any(k in row0_str for k in ["ORIGIN", "FACTOR", "WELDER1"]):
                            self.log(f"   ℹ️ [자동보정] 두 줄 헤더 감지됨 (시트: {sheet_name})")
                            new_cols = []
                            for col, val in zip(df.columns, df.iloc[0]):
                                c_txt = str(col) if pd.notna(col) and "Unnamed" not in str(col) else ""
                                v_txt = str(val) if pd.notna(val) else ""
                                new_cols.append(f"{c_txt} {v_txt}".strip())
                            df.columns = new_cols
                            df = df.iloc[1:].reset_index(drop=True)

                    # --- Column Identification for THIS sheet ---
                    if mode in ["RT", "KOGAS"]:
                        def _get_col(config_key, search_keywords):
                            c_idx_str = str(self.config.get(config_key, "")).strip()
                            if c_idx_str:
                                try:
                                    idx = self.col_to_num(c_idx_str)
                                    if 1 <= idx <= len(df.columns):
                                        candidate = df.columns[idx - 1]
                                        
                                        # [지능형 자동 매핑 우선 정책] 
                                        # 오직 사용자가 설정을 변경하지 않고 '기본값' 상태일 때만, 헤더 매칭이 틀렸을 경우 자동 매핑을 시도합니다.
                                        # 사용자가 명시적으로 컬럼을 직접 변경한 경우는 설정을 100% 신뢰하여 그대로 사용합니다.
                                        DEFAULT_COLUMNS = {
                                            "KOGAS_R_COL_NO": "1", "KOGAS_R_COL_DATE": "2", "KOGAS_R_COL_DWG": "3",
                                            "KOGAS_R_COL_JOINT": "4", "KOGAS_R_COL_LOC": "5", "KOGAS_R_COL_THK": "6",
                                            "KOGAS_R_COL_MAT": "7", "KOGAS_R_COL_SIZE": "8", "KOGAS_R_COL_RES": "9",
                                            "KOGAS_R_COL_WELDER": "10", "KOGAS_R_COL_REM": "11", "KOGAS_R_COL_D1": "12",
                                            "RT_COL_NO": "1", "RT_COL_DATE": "2", "RT_COL_DWG": "3", "RT_COL_JOINT": "4",
                                            "RT_COL_LOC": "5", "RT_COL_THK": "6", "RT_COL_MAT": "7", "RT_COL_SIZE": "0",
                                            "RT_COL_ACC": "8", "RT_COL_REJ": "9", "RT_COL_DEG": "10", "RT_COL_RES": "0",
                                        }
                                        is_default = False
                                        default_val = DEFAULT_COLUMNS.get(config_key)
                                        if default_val is not None:
                                            try:
                                                is_default = (self.col_to_num(c_idx_str) == self.col_to_num(default_val))
                                            except:
                                                is_default = (c_idx_str == default_val)
                                                
                                        if is_default:
                                            cand_up = str(candidate).upper()
                                            if search_keywords and not any(k in cand_up for k in search_keywords):
                                                found = _find_col(df, search_keywords)
                                                if found: return found
                                                
                                        return candidate
                                except: pass
                            return _find_col(df, search_keywords)

                        col_prefix = "KOGAS_R" if mode == "KOGAS" else "RT"
                        col_no = _get_col(f"{col_prefix}_COL_NO", _get_kw('No') or ["NO.", "NO", "SEQ", "ITEM", "순번"])
                        col_dwg = _get_col(f"{col_prefix}_COL_DWG", _get_kw('Dwg') or ["ISO", "DWG", "DRAWING", "도면"])
                        col_joint = _get_col(f"{col_prefix}_COL_JOINT", _get_kw('Joint') or ["JOINT", "WELD NO", "J/N", "FILM IDENT", "용접부"])
                        col_loc = _get_col(f"{col_prefix}_COL_LOC", _get_kw('Loc') or ["LOCATION", "POSITION", "FILM LOC"])
                        col_size = _get_col(f"{col_prefix}_COL_SIZE", _get_kw('Size') or ["SIZE", "DIA", "INCH", "구경", "관경"])
                        col_welder = _get_col(f"{col_prefix}_COL_WELDER", _get_kw('Welder') or ["WELDER", "W/N"])
                        col_remarks = _get_col(f"{col_prefix}_COL_REM", _get_kw('Remarks') or ["REMARKS", "REMARK", "비고"])
                        col_date = _get_col(f"{col_prefix}_COL_DATE", _get_kw('Date') or ["DATE", "검사일"])
                        col_t = _get_col(f"{col_prefix}_COL_THK", _get_kw('T') or ["T", "THICK", "THK", "두께"])
                        col_mat = _get_col(f"{col_prefix}_COL_MAT", _get_kw('Mat') or ["MAT", "MATERIAL"])
                        col_weld = _get_col(f"{col_prefix}_COL_WELD", _get_kw('Weld') or ["WELD", "TYPE"])
                        col_iqi = _get_col(f"{col_prefix}_COL_IQI", _get_kw('IQI') or ["IQI"])
                        col_sens = _get_col(f"{col_prefix}_COL_SENS", _get_kw('Sens') or ["SENS", "SENSITIVITY"])
                        col_den = _get_col(f"{col_prefix}_COL_DEN", _get_kw('Den') or ["DEN", "DENSITY"])
                        col_acc = _get_col(f"{col_prefix}_COL_ACC", _get_kw('Acc') or ["ACC", "합격"])
                        col_rej = _get_col(f"{col_prefix}_COL_REJ", _get_kw('Rej') or ["REJ", "불합격"])
                        col_deg = _get_col(f"{col_prefix}_COL_DEG", _get_kw('Deg') or ["DEG", "물성", "수정브랜드", "GRADE"])
                        col_result = _get_col(f"{col_prefix}_COL_RES", _get_kw('Result') or ["RESULT", "판정"])
                        
                        defect_cols = {}
                        for i in range(1, 16):
                            defect_cols[f"D{i}"] = _get_col(f"{col_prefix}_COL_D{i}", [f"D{i}", f"DEFECT{i}", chr(9311 + i), f"{i}"])
                        defect_cols = {k: v for k, v in defect_cols.items() if v}

                        # [NEW] KOGAS Fallback (B column for Joint, E column for Size) ONLY if not found
                        if mode == "KOGAS":
                            if col_joint is None and len(df.columns) > 1: col_joint = df.columns[1]
                            if col_size is None and len(df.columns) > 4: col_size = df.columns[4]
                    elif mode == "PT":
                        col_no = _find_col(df, ["NO.", "NO", "SEQ", "ITEM"])
                        col_dwg = _find_col(df, ["ISO", "LINE", "DWG", "DRAWING"], exclude=["JOINT", "WELD"]) 
                        col_joint = _find_col(df, ["JOINT NO", "JOINT NUMBER"], exclude=["ISO", "LINE", "ITEM"])
                        if not col_joint:
                            col_joint = _find_col(df, ["JOINT", "WELD"], exclude=["ISO", "LINE", "ITEM", "WPS", "REPORT", "DATE", "TYPE"])
                        col_mat = _find_col(df, ["MAT", "MATERIAL", "재질", "CLASS", "M'TL"])
                        col_size = _find_col(df, ["SIZE", "NPS", "DIA", "INCH", "관경", "ORIGIN", "DI ORIGIN"])
                        col_thk = _find_col(df, ["THK", "THICK", "SCH", "두께"])
                        col_welder = _find_col(df, ["WELDER", "ID", "용접사", "WELDER1"])
                        col_wtype = _find_col(df, ["WELD TYPE", "WELDTYPE", "W.TYPE", "TYPE"], exclude=["JOINT"])
                        col_result = _find_col(df, ["RESULT", "결과", "판정"])
                    # Placeholder for Loc to maintain structure
                        col_loc = None 
                    else:
                        col_cr = _find_col(df, ["CR", "CHROMIUM"]); col_ni = _find_col(df, ["NI", "NICKEL"])
                        col_mo = _find_col(df, ["MO", "MOLYBDENUM"]); col_mn = _find_col(df, ["MN", "MANGANESE"])
                        col_no = _find_col(df, ["NO.", "NO", "SEQ", "NUM", "POS", "ITEM"])
                        col_joint = _find_col(df, ["JOINT", "J/N", "JOINT NO", "PUNCH", "WELD NO"])
                        col_loc = _find_col(df, ["LOCATION", "TEST POSITION", "POINT", "AREA", "POSITION"])
                        col_dwg = _find_col(df, ["ISO", "DWG", "DRAWING", "LINE"])
                        col_grade_orig = _find_col(df, ["GRADE", "MATERIAL", "SPEC", "TYPE"])
                        
                        # [DYNAMIC] Scan for dynamically added filter elements (e.g., C, P, S)
                        custom_cols = {}
                        if hasattr(self, 'element_filters') and self.element_filters:
                            for f_item in self.element_filters:
                                f_key = f_item['key'].get().strip()
                                if f_key and f_key.upper() not in ["CR", "NI", "MO", "MN"]:
                                    c_found = _find_col(df, [f_key.upper()])
                                    if c_found is not None:
                                        custom_cols[f_key] = c_found

                    # [NEW] Forward-fill (Fill Down) for ISO and Joint
                    last_dwg = ""
                    last_joint = ""
                    last_date = ""

                    def clean_v(v):
                        if pd.isna(v): return ""
                        # 만약 날짜/시간(Timestamp) 형태라면 YYYY-MM-DD 형식으로 클리닝
                        if isinstance(v, (datetime.datetime, datetime.date, pd.Timestamp)) or hasattr(v, 'strftime'):
                            try:
                                return v.strftime('%Y-%m-%d')
                            except: pass
                        try:
                            if isinstance(v, (int, float)) or (isinstance(v, str) and v.replace('.','').isdigit()):
                                fv = float(v)
                                if fv.is_integer(): return str(int(fv))
                                return str(fv)
                        except: pass
                        # 만약 문자열이고 날짜 형식인 경우 시간 정보(00:00:00) 제거
                        ret_str = str(v).strip()
                        if len(ret_str) >= 10:
                            if ret_str[4] == '-' and ret_str[7] == '-':
                                ret_str = ret_str[:10]
                            elif ret_str[4] == '.' and ret_str[7] == '.':
                                ret_str = ret_str[:10]
                            elif ret_str[4] == '/' and ret_str[7] == '/':
                                ret_str = ret_str[:10]
                        return ret_str

                if mode == "KOGAS":
                    # ===== 가스공사 모드: 2개 행을 1개 Joint로 페어링 =====
                    num_rows = len(df)
                    r_idx = 0
                    while r_idx < num_rows:
                        row_top = df.iloc[r_idx]
                        row_bot = df.iloc[r_idx + 1] if (r_idx + 1 < num_rows) else None

                        # 1. 순번 (No)
                        v_raw_no = clean_v(row_top[col_no]) if col_no is not None else ""
                        if not v_raw_no and row_bot is not None and col_no is not None:
                            v_raw_no = clean_v(row_bot[col_no])
                        
                        # 만약 NO 컬럼이 지정되어 있는데 값이 완전히 비어있거나 'nan'인 경우, 데이터가 없는 행이므로 건너뜁니다.
                        if col_no is not None and (not v_raw_no or v_raw_no == 'nan'):
                            r_idx += 2
                            continue
                            
                        # NO 컬럼이 지정되어 있지 않은 경우에만 인덱스 기반으로 번호를 생성합니다.
                        if not v_raw_no or v_raw_no == 'nan':
                            v_raw_no = str((r_idx // 2) + 1)

                        if target_no_list and v_raw_no not in target_no_list:
                            r_idx += 2
                            continue

                        # 키워드 필터링
                        extract_key = self.rt_extract_keyword.get().strip().lower()
                        if extract_key:
                            top_vals = [clean_v(v).lower() for v in row_top.values if pd.notna(v)]
                            bot_vals = [clean_v(v).lower() for v in row_bot.values if pd.notna(v)] if row_bot is not None else []
                            row_str = ' '.join(top_vals + bot_vals)
                            if extract_key not in row_str:
                                r_idx += 2
                                continue

                        # 2. 도면번호 (Dwg) 및 Dwg_Sub
                        curr_dwg = sheet_level_dwg if sheet_level_dwg else (str(row_top[col_dwg]).strip() if col_dwg is not None else '')
                        if (not curr_dwg or curr_dwg == 'nan'):
                            curr_dwg = last_dwg if last_dwg else ''
                        if curr_dwg and curr_dwg != 'nan': last_dwg = curr_dwg

                        curr_dwg_sub = ''
                        if sheet_level_dwg:
                            curr_dwg_sub = sheet_level_dwg
                        elif row_bot is not None and col_dwg is not None:
                            curr_dwg_sub = str(row_bot[col_dwg]).strip()
                        if not curr_dwg_sub or curr_dwg_sub == 'nan':
                            curr_dwg_sub = curr_dwg

                        # 3. 공동 (Joint)
                        curr_joint = str(row_top[col_joint]).strip() if col_joint is not None else ''
                        if (not curr_joint or curr_joint == 'nan') and last_joint: curr_joint = last_joint
                        if curr_joint and curr_joint != 'nan': last_joint = curr_joint
                        elif not curr_joint or curr_joint == 'nan': curr_joint = v_raw_no

                        # 4. 일자 (Date)
                        curr_date = clean_v(row_top[col_date]) if col_date is not None else ''
                        if (not curr_date or curr_date.lower() == 'nan') and last_date: curr_date = last_date
                        if curr_date and curr_date.lower() != 'nan': last_date = curr_date
                        elif not curr_date or curr_date.lower() == 'nan': curr_date = sheet_level_date if sheet_level_date else ''

                        # 5. 규격 (Size) 및 촬영 매수 (Shots)
                        raw_size = str(row_top[col_size]).strip() if col_size is not None else ''
                        s_match = re.search(r'(\d+\.?\d*)', raw_size)
                        size_val = s_match.group(1) if s_match else raw_size
                        num_shots = self.calculate_rt_shots(size_val)

                        # 6. 용접사 (Welder) 및 Welder_Sub
                        raw_welder = str(row_top[col_welder]).strip() if col_welder is not None else ''
                        v_welder = raw_welder[-3:] if len(raw_welder) > 3 else raw_welder

                        raw_welder_sub = ''
                        if row_bot is not None and col_welder is not None:
                            raw_welder_sub = str(row_bot[col_welder]).strip()
                        v_welder_sub = raw_welder_sub[-3:] if len(raw_welder_sub) > 3 else raw_welder_sub
                        if not v_welder_sub or v_welder_sub == 'nan':
                            v_welder_sub = v_welder

                        # 7. 재질 (Mat) 및 Mat_Sub
                        raw_mat = clean_v(row_top[col_mat]) if col_mat is not None else ''
                        raw_mat_sub = ''
                        if row_bot is not None and col_mat is not None:
                            raw_mat_sub = clean_v(row_bot[col_mat])
                        if not raw_mat_sub or raw_mat_sub == 'nan':
                            raw_mat_sub = raw_mat

                        # 8. 두께 (T)
                        raw_t = clean_v(row_top[col_t]) if col_t is not None else ''

                        # 9. 위치 (Loc) - 촬영 위치 추출 추가!
                        v_loc = clean_v(row_top[col_loc]) if col_loc is not None else ''
                        if not v_loc or v_loc == 'nan':
                            v_loc = '-'

                        item_data = {
                            'No': v_raw_no, 'Date': curr_date,
                            'Dwg': curr_dwg, 'Dwg_Sub': curr_dwg_sub,
                            'Joint': curr_joint, 'Loc': v_loc,
                            'Acc': clean_v(row_top[col_acc]) if col_acc is not None else '',
                            'Rej': clean_v(row_top[col_rej]) if col_rej is not None else '',
                            'Deg': clean_v(row_top[col_deg]) if col_deg is not None else '',
                            'Welder': v_welder, 'Welder_Sub': v_welder_sub,
                            'Remarks': clean_v(row_top[col_remarks]) if col_remarks is not None else '',
                            'T': raw_t,
                            'Mat': raw_mat, 'Mat_Sub': raw_mat_sub,
                            'Weld': clean_v(row_top[col_weld]) if col_weld is not None else '',
                            'IQI': clean_v(row_top[col_iqi]) if col_iqi is not None else '',
                            'Sens': clean_v(row_top[col_sens]) if col_sens is not None else '',
                            'Den': clean_v(row_top[col_den]) if col_den is not None else '',
                            'Result': clean_v(row_top[col_result]) if col_result is not None else 'ACC',
                            'Size': size_val, 'num_shots': num_shots,
                            'selected': True, 'order_index': len(self.rt_extracted_data) + len(all_extracted_data),
                            '_src': {
                                'sheet': sheet_name,
                                'row': r_idx + header_idx + 2,
                                'col_result': list(df.columns).index(col_result) + 1 if col_result is not None else None,
                                'col_remarks': list(df.columns).index(col_remarks) + 1 if col_remarks is not None else None,
                                'col_acc': list(df.columns).index(col_acc) + 1 if col_acc is not None else None,
                                'col_rej': list(df.columns).index(col_rej) + 1 if col_rej is not None else None,
                                'col_defects': {k: list(df.columns).index(v) + 1 for k, v in defect_cols.items()}
                            }
                        }
                        # Defects
                        for i in range(1, 16):
                            key = f'D{i}'; src_col = defect_cols.get(key)
                            src_val = str(row_top[src_col]).strip() if src_col is not None else ''
                            item_data[key] = src_val if src_val else ('1' if i <= num_shots else '')

                        all_extracted_data.append(item_data)
                        r_idx += 2

                else:
                    # ===== 표준 모드 (RT, PT, PMI) =====
                    for _, row in df.iterrows():
                        v_raw_no = clean_v(row[col_no]) if col_no is not None else str(_+1)
                        if col_no is not None and (not v_raw_no or v_raw_no == "nan"):
                            continue
                        if target_no_list and v_raw_no not in target_no_list: continue

                        extract_key = self.rt_extract_keyword.get().strip().lower()
                        if extract_key:
                            row_vals = [clean_v(v).lower() for v in row.values if pd.notna(v)]
                            row_str = ' '.join(row_vals)
                            if extract_key not in row_str:
                                continue

                        curr_dwg = sheet_level_dwg if sheet_level_dwg else (str(row[col_dwg]).strip() if col_dwg is not None else '')
                        if (not curr_dwg or curr_dwg == 'nan'):
                            curr_dwg = last_dwg if last_dwg else ''
                        if curr_dwg and curr_dwg != 'nan': last_dwg = curr_dwg

                        curr_joint = str(row[col_joint]).strip() if col_joint is not None else ''
                        if (not curr_joint or curr_joint == 'nan') and last_joint: curr_joint = last_joint
                        if curr_joint and curr_joint != 'nan': last_joint = curr_joint
                        elif not curr_joint or curr_joint == 'nan': curr_joint = v_raw_no

                        curr_date = clean_v(row[col_date]) if col_date is not None else ''
                        if (not curr_date or curr_date.lower() == 'nan') and last_date: curr_date = last_date
                        if curr_date and curr_date.lower() != 'nan': last_date = curr_date
                        elif not curr_date or curr_date.lower() == 'nan': curr_date = sheet_level_date if sheet_level_date else ''

                        if mode == 'RT':
                            raw_size = str(row[col_size]).strip() if col_size is not None else ''
                            s_match = re.search(r'(\d+\.?\d*)', raw_size)
                            size_val = s_match.group(1) if s_match else raw_size

                            num_shots = self.calculate_rt_shots(size_val)

                            raw_welder = str(row[col_welder]).strip() if col_welder is not None else ''
                            v_welder = raw_welder[-3:] if len(raw_welder) > 3 else raw_welder

                            v_date = curr_date

                            item_data = {
                                'No': v_raw_no, 'Date': v_date,
                                'Dwg': curr_dwg, 'Joint': curr_joint, 'Loc': '-',
                                'Acc': clean_v(row[col_acc]) if col_acc is not None else '',
                                'Rej': clean_v(row[col_rej]) if col_rej is not None else '',
                                'Deg': clean_v(row[col_deg]) if col_deg is not None else '',
                                'Welder': v_welder,
                                'Remarks': clean_v(row[col_remarks]) if col_remarks is not None else '',
                                'T': clean_v(row[col_t]) if col_t is not None else '',
                                'Mat': clean_v(row[col_mat]) if col_mat is not None else '',
                                'Weld': clean_v(row[col_weld]) if col_weld is not None else '',
                                'IQI': clean_v(row[col_iqi]) if col_iqi is not None else '',
                                'Sens': clean_v(row[col_sens]) if col_sens is not None else '',
                                'Den': clean_v(row[col_den]) if col_den is not None else '',
                                'Result': clean_v(row[col_result]) if col_result is not None else 'ACC',
                                'Size': size_val, 'num_shots': num_shots,
                                'selected': True, 'order_index': len(self.rt_extracted_data) + len(all_extracted_data),
                                '_src': {
                                    'sheet': sheet_name,
                                    'row': _ + header_idx + 2,
                                    'col_result': list(df.columns).index(col_result) + 1 if col_result is not None else None,
                                    'col_remarks': list(df.columns).index(col_remarks) + 1 if col_remarks is not None else None,
                                    'col_acc': list(df.columns).index(col_acc) + 1 if col_acc is not None else None,
                                    'col_rej': list(df.columns).index(col_rej) + 1 if col_rej is not None else None,
                                    'col_defects': {k: list(df.columns).index(v) + 1 for k, v in defect_cols.items()}
                                }
                            }
                            for i in range(1, 16):
                                key = f'D{i}'; src_col = defect_cols.get(key)
                                src_val = str(row[src_col]).strip() if src_col is not None else ''
                                item_data[key] = src_val if src_val else ('1' if i <= num_shots else '')

                            if num_shots > 1:
                                for shot_idx in range(1, num_shots + 1):
                                    shot_item = item_data.copy()
                                    shot_item['Loc'] = f'{shot_idx}-{(shot_idx % num_shots) + 1}'
                                    for i in range(1, 16):
                                        key = f'D{i}'; c = defect_cols.get(key); val = str(row[c]).strip() if c is not None else ''
                                        shot_item[key] = '?' if val and val.lower() in ['v', 'x', 'o', '1', '?'] else ''
                                    all_extracted_data.append(shot_item)
                            else:
                                all_extracted_data.append(item_data)
                        elif mode == "PT":
                            res_str = str(row[col_result]).upper() if col_result is not None else "ACC"
                            if any(k in res_str for k in ["ACC", "OK", "ACCEPT", "합격"]):
                                item_row = {
                                    'No': v_raw_no, 'Date': sheet_level_date, 'Dwg': curr_dwg, 'Joint': self.force_two_digit(curr_joint),
                                    'NPS': str(row[col_size]).strip() if col_size is not None else "", 'Thk.': convert_sch_to_thk(str(row[col_size]).strip() if col_size is not None else "", str(row[col_thk]).strip() if col_thk is not None else ""),
                                    'Material': self.fix_material_name(row[col_mat]) if col_mat is not None else "", 'Welder': str(row[col_welder]).strip() if col_welder is not None else "",
                                    'WType': str(row[col_wtype]).strip() if col_wtype is not None else "", 'Result': "Acc", 'selected': True,
                                    'order_index': len(self.pt_extracted_data) + len(all_extracted_data)
                                }
                                for k in self.pt_column_keys:
                                    if k not in item_row and k != "selected": item_row[k] = ""
                                all_extracted_data.append(item_row)
                        else:
                            # PMI-specific Row Data
                            v_cr = self.to_float(row[col_cr])
                            if v_cr > 0 or (v_raw_no != "" and v_raw_no != "nan"):
                                v_ni = self.to_float(row[col_ni])
                                v_mo = self.to_float(row[col_mo]) if col_mo is not None else 0.0
                                v_mn = self.to_float(row[col_mn]) if col_mn is not None else 0.0
                                orig_grade = str(row[col_grade_orig]).strip() if col_grade_orig is not None else ""
                                
                                final_grade = orig_grade
                                if self.auto_verify.get():
                                    detected = self.check_material_grade({'Cr': v_cr, 'Ni': v_ni, 'Mo': v_mo, 'Mn': v_mn})
                                    if detected: final_grade = detected
                                if not final_grade or final_grade == "nan":
                                    final_grade = "SS316" if v_mo >= 1.5 else "SS304"

                                pmi_item = {
                                    'No': v_raw_no, 
                                    'Joint': curr_joint,
                                    'Loc': str(row[col_loc]).strip() if col_loc is not None else "",
                                    'Cr': v_cr, 'Ni': v_ni, 'Mo': v_mo, 'Mn': v_mn,
                                    'Grade': final_grade, 'Dwg': curr_dwg,
                                    'Date': sheet_level_date,
                                    'selected': True,
                                    'order_index': len(self.extracted_data) + len(all_extracted_data)
                                }
                                # Extract dynamically found elements
                                for c_key, c_idx in custom_cols.items():
                                    pmi_item[c_key] = self.to_float(row[c_idx])
                                # [NEW] 기존 수동 추가 컬럼 보존
                                for k in self.column_keys:
                                    if k not in pmi_item and k != "selected":
                                        pmi_item[k] = ""
                                all_extracted_data.append(pmi_item)
                self.progress['value'] = ((s_idx + 1) / len(xls.sheet_names)) * 50

            if not all_extracted_data:
                messagebox.showerror("오류", "추출된 데이터가 없습니다.")
                return False

            # Extraction Mode Filter (PMI Only)
            if mode == "PMI":
                ext_mode = self.extraction_mode.get()
                if ext_mode != "전체":
                    original_count = len(all_extracted_data)
                    if ext_mode == "SS304 만": all_extracted_data = [d for d in all_extracted_data if d['Grade'] == "SS304"]
                    elif ext_mode == "SS316 만": all_extracted_data = [d for d in all_extracted_data if d['Grade'] == "SS316"]
                    elif ext_mode == "DUPLEX 만": all_extracted_data = [d for d in all_extracted_data if d['Grade'] == "DUPLEX"]
                    elif ext_mode == "SS310 만": all_extracted_data = [d for d in all_extracted_data if d['Grade'] == "SS310"]
                    elif ext_mode == "미분류(기타) 만":
                        known_grades = ["SS304", "SS316", "DUPLEX", "SS310"]
                        all_extracted_data = [d for d in all_extracted_data if d['Grade'] not in known_grades]
                    
                    self.log(f"🔍 필터링 ({ext_mode}): {original_count}개 -> {len(all_extracted_data)}개")
                    if not all_extracted_data:
                        messagebox.showinfo("알림", f"'{ext_mode}'에 해당하는 데이터가 없습니다.")
                        return False

            # [NEW] 원소 함량 필터링 적용 (PMI 전용)
            if mode == "PMI" and hasattr(self, 'element_filters') and self.element_filters:
                for f_item in self.element_filters:
                    f_key = f_item['key'].get().strip()
                    if not f_key: continue
                    f_min = self.to_float(f_item['min'].get())
                    f_max = self.to_float(f_item['max'].get())
                    
                    if f_min > 0 or f_max > 0:
                        original_count = len(all_extracted_data)
                        all_extracted_data = [d for d in all_extracted_data if (f_min <= d.get(f_key, 0.0) <= (f_max if f_max > 0 else 999.9))]
                        if len(all_extracted_data) != original_count:
                             self.log(f"🔍 원소 필터 적용 ({f_key}: {f_min}~{f_max}): {original_count} -> {len(all_extracted_data)}건 남음")

            if mode == "PT":
                # [NEW] [Refinement] Duplicate removal for PT (ISO + Joint)
                seen = set()
                unique_data = []
                for d in all_extracted_data:
                    key = (str(d['Dwg']).strip(), str(d['Joint']).strip())
                    if key not in seen:
                        seen.add(key)
                        unique_data.append(d)
                all_extracted_data = unique_data

            if target_no_list:
                all_extracted_data.sort(key=lambda x: target_no_list.index(str(x['No'])) if str(x['No']) in target_no_list else 999999)

            # [CHANGE] Overwrite -> Accumulate + Auto Sort
            if mode == "RT":
                self.rt_extracted_data.extend(all_extracted_data)
                self.update_date_listbox("RT")
                self._update_rt_preview_columns(mode="RT") # [FIX] 표시 이름 보존을 위해 호출
                self.sort_by_column("Dwg", mode="RT") 
                total_count = len(self.rt_extracted_data)
            elif mode == "KOGAS":
                self.kogas_extracted_data.extend(all_extracted_data)
                self.update_date_listbox("KOGAS")
                self._update_rt_preview_columns(mode="KOGAS") # [FIX] 표시 이름 보존을 위해 호출
                self.sort_by_column("Dwg", mode="KOGAS")
                total_count = len(self.kogas_extracted_data)
            elif mode == "PT":
                self.pt_extracted_data.extend(all_extracted_data)
                self.update_date_listbox("PT")
                self.sort_by_column("Dwg", mode="PT") 
                total_count = len(self.pt_extracted_data)
            else:
                self.extracted_data.extend(all_extracted_data)
                self.update_date_listbox("PMI")
                self.sort_by_column("Dwg", mode="PMI")
                total_count = len(self.extracted_data)
            
            self.progress['value'] = 100
            if show_msg:
                self.log(f"✅ {mode} 데이터 누적 완료 (현재 총 {total_count} 건)")
                messagebox.showinfo("완료", f"데이터가 추가되었습니다.\n현재 목록에는 총 {total_count}건의 데이터가 있습니다.")
            return True
        except Exception as e:
            self.log(f"❌ {mode} 추출 오류: {e}")
            traceback.print_exc()
            return False

    def sync_results_to_request(self):
        """입력된 결과를 원본 의뢰서 파일로 역전송 (RT/KOGAS 대응)"""
        try:
            main_tab = self.mode_notebook.tab(self.mode_notebook.select(), "text")
            if "RT" in main_tab:
                sub_tab = self.rt_preview_nb.tab(self.rt_preview_nb.select(), "text")
                mode = "KOGAS" if "가스공사" in sub_tab else "RT"
            else:
                messagebox.showwarning("경고", "현재 RT 모드에서만 지원됩니다.")
                return

            target_file = self.rt_target_file_path.get() if mode == "RT" else self.kogas_target_file_path.get()
            if not target_file or not os.path.exists(target_file):
                messagebox.showwarning("파일 오류", "원본 의뢰서 파일을 찾을 수 없습니다.")
                return

            data = self.rt_extracted_data if mode == "RT" else self.kogas_extracted_data
            if not data:
                messagebox.showwarning("데이터 오류", "반영할 데이터가 없습니다.")
                return

            if not messagebox.askyesno("확인", f"현재 입력된 결과를 원본 의뢰서({os.path.basename(target_file)})에 반영하시겠습니까?"):
                return

            self.log(f"🔄 의뢰서 결과 반영 시작: {os.path.basename(target_file)}")
            
            import openpyxl
            wb = openpyxl.load_workbook(target_file)
            count = 0
            
            # Group items by source row to handle multi-shot joints
            groups = {}
            for item in data:
                if not item.get('_src'): continue
                src_key = (item['_src']['sheet'], item['_src']['row'])
                if src_key not in groups: groups[src_key] = []
                groups[src_key].append(item)

            for (sheet_name, row_idx), items in groups.items():
                if sheet_name not in wb.sheetnames: continue
                ws = wb[sheet_name]
                src = items[0]['_src']
                
                # Determine final result (If any item is Rej, result is Rej)
                final_res = "ACC"
                all_remarks = []
                has_defects = {k: False for k in src.get('col_defects', {}).keys()}
                
                for it in items:
                    res = str(it.get('Result', '')).upper()
                    if any(x in res for x in ["REJ", "NG", "불합격"]): final_res = "REJ"
                    
                    rem = str(it.get('Remarks', '')).strip()
                    if rem and rem not in all_remarks: all_remarks.append(rem)
                    
                    if src.get('col_defects'):
                        for d_key in src['col_defects'].keys():
                            if it.get(d_key) in ["√", "1", "v", "V", "o", "O"]: has_defects[d_key] = True

                # Write back to Excel
                if src.get('col_result'):
                    ws.cell(row=row_idx, column=src['col_result'], value=final_res)
                
                if src.get('col_remarks') and all_remarks:
                    ws.cell(row=row_idx, column=src['col_remarks'], value=", ".join(all_remarks))
                
                if src.get('col_defects'):
                    for d_key, col_idx in src['col_defects'].items():
                        if has_defects[d_key]:
                            ws.cell(row=row_idx, column=col_idx, value="√")
                
                count += 1

            wb.save(target_file)
            self.log(f"✅ 총 {count}개 행의 결과를 의뢰서에 성공적으로 반영했습니다.")
            messagebox.showinfo("완료", f"{count}개 행의 결과가 의뢰서에 반영되었습니다.")

        except Exception as e:
            self.log(f"❌ 의뢰서 결과 반영 실패: {e}")
            messagebox.showerror("오류", f"의뢰서 결과 반영 중 오류가 발생했습니다: {e}")

    def duplicate_selected_rows(self):
        """선택된 행들을 복제하여 데이터 리스트에 삽입"""
        selected = self.preview_tree.selection()
        if not selected: return
        
        # 시각적 순서대로 정렬 (위에서부터 복제)
        selected = sorted(list(selected), key=lambda x: self.preview_tree.index(x))
        
        new_items = []
        insert_pos = 0
        for item_id in selected:
            view_idx = self.preview_tree.index(item_id)
            actual_idx = self.item_idx_map[view_idx]
            new_item = self.extracted_data[actual_idx].copy()
            # No.는 중복 방지를 위해 비우거나 새로 부여 (여기서는 유지 후 나중에 rename 가능)
            new_items.append(new_item)
            insert_pos = actual_idx
            
        for item in reversed(new_items):
            self.extracted_data.insert(insert_pos + 1, item)
             
        self.populate_preview(self.extracted_data, switch_tab=False)
        self.log(f"👯 {len(new_items)}개 행 복제 완료")

    def save_column_widths(self):
        """현재 트리뷰의 컬럼 너비를 설정 파일에 저장"""
        widths = {}
        for col in self.preview_tree["columns"]:
            widths[col] = self.preview_tree.column(col, "width")
        
        self.config["PMI_COL_WIDTHS"] = widths
        self.save_settings()
        self.log("📏 컬럼 너비 설정 저장 완료")

    def apply_date_filter(self):
        """날짜 리스트박스 선택 상태를 데이터 모델에 반영"""
        selected_dates = []
        for i in range(self.date_listbox.size()):
            val = self.date_listbox.get(i)
            if val.startswith("[v]"):
                selected_dates.append(val.replace("[v] ", "").strip())
        
        if not selected_dates:
            messagebox.showwarning("필터 오류", "최소 하나 이상의 날짜를 선택해주세요.")
            return

        for item in self.extracted_data:
            item['date_filtered'] = (item.get('Date', '') in selected_dates)
        
        self.populate_preview(self.extracted_data, switch_tab=False)
        self.log(f"📅 날짜 필터 적용 완료: {len(selected_dates)}개 날짜 선택됨")

    def run_process(self):
        # [NEW] Ensure config values are correct types for comparison (prevent str vs int errors)
        for k in list(self.config.keys()):
            if k.endswith(('_ROW', '_IDX', '_SIZE')) or any(x in k for x in ['START', 'END', 'PAGE']):
                try: self.config[k] = int(float(self.config[k]))
                except: pass
            elif any(x in k for x in ['MARGIN', 'SCALE', 'RATIO', 'POS', 'HEIGHT', 'WIDTH']):
                try: self.config[k] = float(self.config[k])
                except: pass

        # 현재 활성 탭에 따른 모드 결정
        try:
            tab_text = self.mode_notebook.tab(self.mode_notebook.select(), "text")
            if "RT" in tab_text:
                sub_tab = self.rt_preview_nb.tab(self.rt_preview_nb.select(), "text")
                mode = "KOGAS" if "가스공사" in sub_tab else "RT"
            elif "PT" in tab_text: mode = "PT"
            elif "PAUT" in tab_text: mode = "PAUT"
            else: mode = "PMI"
        except: mode = "PMI"

        if mode == "RT":
            target_file = self.rt_target_file_path.get()
            template_path = self.rt_template_file_path.get()
            data = self.rt_extracted_data
        elif mode == "KOGAS":
            target_file = self.kogas_target_file_path.get()
            template_path = self.kogas_template_file_path.get()
            data = self.kogas_extracted_data
        elif mode == "PT":
            target_file = self.pt_target_file_path.get()
            template_path = self.pt_template_file_path.get()
            data = self.pt_extracted_data
        else:
            target_file = self.target_file_path.get()
            template_path = self.template_file_path.get()
            data = self.extracted_data
        
        if not template_path:
            messagebox.showwarning("파일 미선택", f"{mode} 양식(Template) 파일을 선택해주세요.")
            return
            
        if not target_file and not data:
            messagebox.showwarning("파일 미선택", f"{mode} 데이터 파일(Excel)을 선택하거나, 저장된 데이터를 불러와주세요.")
            return

        # [NEW] 템플릿 파일 존재 여부 및 기본 구조 확인
        if not os.path.exists(template_path):
            messagebox.showerror("오류", f"템플릿 파일을 찾을 수 없습니다:\n{template_path}")
            return
        self.save_settings() # This now captures all setting_vars (Logo X/Y, Print Area, etc.)
        
        # 데이터가 비어있거나 새로 추출이 필요한 경우 수행
        if not data:
            if not self.extract_only(show_msg=False): return
            # Re-fetch data after extraction
            if mode == "PT": data = self.pt_extracted_data
            elif mode == "RT": data = self.rt_extracted_data
            elif mode == "KOGAS": data = self.kogas_extracted_data
            else: data = self.extracted_data
            
        # [NEW] 체크된 항목만 필터링 (기본값은 True)
        final_list = [d for d in data if d.get('selected', True) and d.get('date_filtered', True)]
        if not final_list:
            messagebox.showwarning("항목 미선택", f"선택된 {mode} 데이터가 없습니다. 미리보기에서 항목을 체크해주세요.")
            return

        if mode == "RT":
            self._run_rt_process(final_list, template_path, mode="RT")
        elif mode == "KOGAS":
            self._run_rt_process(final_list, template_path, mode="KOGAS")
        elif mode == "PT":
            self._run_pt_process(final_list, template_path)
        else:
            self._run_pmi_process(final_list, template_path)

    def _write_gapji_metadata(self, ws):
        """Write common report metadata to the cover sheet."""
        # Default mapping: Project: B5, Customer: B6, Item: B7, Material: B8, Date: B9, Report No: B10
        mapping = [
            ('GAPJI_PROJECT', 'B5'),
            ('GAPJI_CUSTOMER', 'B6'),
            ('GAPJI_ITEM', 'B7'),
            ('GAPJI_MATERIAL', 'B8'),
            ('GAPJI_EXAM_DATE', 'B9'),
            ('GAPJI_REPORT_NO', 'B10')
        ]
        
        for cfg_key, coord in mapping:
            val = self.config.get(cfg_key, "")
            if val:
                self.safe_set_value(ws, coord, val)

    def _run_pmi_process(self, final_list, template_path):
        self.save_settings() # Ensure UI -> config sync
        self.log(f"🚀 PMI 성적서 생성 시작 (총 {len(final_list)} 건)...")
        self.progress['value'] = 0
        all_extracted_data = final_list
        
        # [FIX] data_end_row 전역/설정으로 보강
        data_start_row = int(self.config.get('START_ROW', 17))
        data_end_row = int(self.config.get('DATA_END_ROW', 45))
        
        try:
            wb = openpyxl.load_workbook(template_path)
            if len(wb.worksheets) < 1:
                raise ValueError("선택한 템플릿 파일에 시트가 존재하지 않습니다.")

            if len(wb.worksheets) >= 1:
                ws0 = wb.worksheets[0]
                self.add_logos_to_sheet(ws0, is_cover=True, clear_existing=False, mode="PMI")
                self._write_gapji_metadata(ws0)
                self.force_print_settings(ws0, context="COVER") # [NEW] 갑지 전용 여백 적용
                
                # [DYNAMIC ELEMENTS] Gapji Signature Area Border Stabilization
                # Signature usually ends at row 48. We must ensure no vertical lines extend below it.
                b_start = int(self.config.get('GAPJI_START_ROW', 23))
                b_end = int(self.config.get('GAPJI_DATA_END_ROW', 38))
                sig_end = 48 # Standard signature end row for this template
                
                # 1. Apply side borders only within the valid range
                for r in range(b_start, sig_end + 1):
                    try:
                        # Left Border (Column A)
                        cell_a = ws0.cell(row=r, column=1); eb = cell_a.border
                        ws0.cell(row=r, column=1).border = Border(left=medium_side, right=eb.right, top=eb.top, bottom=eb.bottom)
                        # Right Border (Column M/13)
                        cell_m = ws0.cell(row=r, column=13); eb_m = cell_m.border
                        ws0.cell(row=r, column=13).border = Border(left=eb_m.left, right=medium_side, top=eb_m.top, bottom=eb_m.bottom)
                    except: pass
                
                # 2. [FIX] AGGRESSIVE CLEAR below row 48 (Signature end)
                # Unmerge and clear all borders to ensure no ghost lines remain
                for r in range(sig_end + 1, sig_end + 30):
                    # Clear merges first
                    merged_ranges = [str(rng) for rng in ws0.merged_cells.ranges if rng.min_row >= r]
                    for m_rng in merged_ranges:
                        try: ws0.unmerge_cells(m_rng)
                        except: pass
                    # Clear all column borders in this row
                    for c in range(1, 26): # Columns A to Z
                        try: ws0.cell(row=r, column=c).border = Border()
                        except: pass

                # [FORCE] Strict Print Area for Gapji to prevent printing row 49+
                ws0.print_area = f'A1:M{sig_end}'
                
                ws0['I35'].border = Border() # [FIX] I35 셀 선 제거
                self.safe_set_value(ws0, 'I35', None) 
                self.apply_custom_dimensions(ws0, "COVER")
            
            data_sheet_id = 1 if len(wb.worksheets) >= 2 else 0
            ws = wb.worksheets[data_sheet_id]; ws.title = f"{ws.title[:20]}_001"
            # 을지 기본 설정
            self.add_logos_to_sheet(ws, is_cover=False, clear_existing=(ws != ws0), mode="PMI")
            self.force_print_settings(ws, context="DATA"); self.set_eulji_headers(ws)
            # self.apply_custom_dimensions(ws, "DATA") # [MOVED] To the end of process
            
            try:
                if len(wb.worksheets) >= 2:
                    # [FIX] 수식으로 연결 (갑지 L5 -> 을지 K5, 갑지 N5 -> 을지 M5, 갑지 N8 -> 을지 M8)
                    self.safe_set_value(ws, 'K5', f"='{ws0.title}'!L5")
                    self.safe_set_value(ws, 'M5', f"='{ws0.title}'!N5")
                    self.safe_set_value(ws, 'M8', f"='{ws0.title}'!N8")
                    # [FIX] K5:M10 범위 스타일 설정 (바탕, 크기 9)
                    for r_idx in range(5, 11):
                        for c_idx in range(11, 14):
                            cell = ws.cell(row=r_idx, column=c_idx)
                            cell.font = Font(name='바탕', size=9, bold=False)
                            if (r_idx == 5 or r_idx == 8) and c_idx == 11: # K5, K8 줄바꿈
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            else:
                                cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                    ws.column_dimensions['M'].width = 18.0
            except Exception as e:
                self.log(f"갑지 데이터 복사 실패: {e}")
            
            materials = "SS304,SS304L,SS316,SS316L,SS321,SS347,SS410,SS430,DUPLEX,MONEL,INCONEL,ER308,ER308L,ER309,ER309L,ER316,ER316L,ER347,ER2209,WP316,WP316L,TP316,TP316L,F316L,A182-F316L,A312-TP316L"
            dv_q = DataValidation(type="list", formula1=f'"{materials}"', allow_blank=True)
            # [NEW] 데이터 기입 전 기존 병합 영역 정리 (A-E 등 중복 병합 방지)
            def clear_merges_in_range(sheet, start_row, end_row):
                merged_ranges = list(sheet.merged_cells.ranges)
                for r in merged_ranges:
                    # 영역이 겹치면 해제 (aggressive clearing)
                    if r.min_row <= end_row and r.max_row >= start_row:
                        try: sheet.unmerge_cells(str(r))
                        except: pass

            ws = wb.worksheets[data_sheet_id]
            clear_merges_in_range(ws, self.config['START_ROW'], self.config['DATA_END_ROW'] + 20)
            ws.add_data_validation(dv_q)

            # [DYNAMIC ELEMENTS] Identify newly appended chemical elements
            custom_elements = [k for k in self.column_keys if k not in ["_status", "selected", "No", "Date", "Dwg", "Joint", "Loc", "Grade", "order_index", "V", "ST", "Ni", "Cr", "Mo"]]
            
            # [NEW] Dynamic Column Mapping for Elements
            ni_col = self.col_to_num(self.config.get('PMI_COL_NI', '8'))
            cr_col = self.col_to_num(self.config.get('PMI_COL_CR', '9'))
            mo_col = self.col_to_num(self.config.get('PMI_COL_MO', '10'))
            
            print_elements = [('Ni', ni_col), ('Cr', cr_col), ('Mo', mo_col)]
            if len(custom_elements) > 0: print_elements.append((custom_elements[0], 11))
            if len(custom_elements) > 1: print_elements.append((custom_elements[1], 12))

            current_row = self.config['START_ROW']; current_page = 1; data_ptr = 0
            while data_ptr < len(all_extracted_data):
                # 가용 행 수 계산 (DATA_END_ROW까지만 채움)
                rows_left = self.config['DATA_END_ROW'] - current_row + 1
                
                # 만약 공간이 전혀 없으면 새 시트로 전환
                if rows_left <= 0:
                    current_page += 1; ws = self.prepare_next_sheet(wb, data_sheet_id, current_page)
                    clear_merges_in_range(ws, self.config['START_ROW'], self.config['DATA_END_ROW'] + 20)
                    current_row = self.config['START_ROW']; ws.add_data_validation(dv_q)
                    rows_left = self.config['DATA_END_ROW'] - current_row + 1

                # [NEW] Dynamically print headers for the new elements precisely on the first row of each sheet
                if current_row == self.config['START_ROW'] and len(print_elements) > 3:
                    h_row = self.config['START_ROW'] - 1
                    for r_i in range(max(1, self.config['START_ROW'] - 5), self.config['START_ROW']):
                        if str(ws.cell(row=r_i, column=8).value).strip().upper() == "NI":
                            h_row = r_i; break
                    for val_k, c_idx in print_elements[3:]:
                        hdr_cell = ws.cell(row=h_row, column=c_idx)
                        self.safe_set_value(ws, hdr_cell.coordinate, val_k, align='center')
                        hdr_cell.font = Font(name=hdr_cell.font.name if hasattr(hdr_cell.font, 'name') else 'Arial', size=hdr_cell.font.size if hasattr(hdr_cell.font, 'size') else 9, bold=True)

                # [NEW DYNAMIC ALGORITHM] Determine block size based on the UI's manual merge logic
                # We start with the item at data_ptr. We scan forward to see how many subsequent items
                # legitimately belong to this same visual block.
                logical_block_size = 1
                base_item = all_extracted_data[data_ptr]
                base_joint = base_item.get('visual_group_joint', base_item.get('Joint', ''))
                base_iso = base_item.get('visual_group_iso', base_item.get('ISO', base_item.get('Dwg', '')))
                
                # Check consecutive items
                for ahead_idx in range(data_ptr + 1, min(len(all_extracted_data), data_ptr + rows_left)):
                    ahead_item = all_extracted_data[ahead_idx]
                    
                    # Verify Joint Merge Configuration
                    ui_merged_j = ahead_item.get('is_merged_joint') == True
                    name_merged_j = ahead_item.get('visual_group_joint', ahead_item.get('Joint', '')) == base_joint
                    joint_ok = ui_merged_j or name_merged_j

                    # Verify ISO Merge Configuration
                    ui_merged_i = ahead_item.get('is_merged_iso') == True
                    name_merged_i = ahead_item.get('visual_group_iso', ahead_item.get('ISO', ahead_item.get('Dwg', ''))) == base_iso
                    iso_ok = ui_merged_i or name_merged_i
                    
                    # If BOTH are ok, it's a block. If EITHER is broken (by User or Data mismatch), we split the block!
                    if joint_ok and iso_ok:
                        logical_block_size += 1
                    else:
                        break
                
                # Cap the block size strictly to the available rows remaining on the sheet
                this_block_size = min(logical_block_size, rows_left)
                actual_block_rows = this_block_size
                batch = all_extracted_data[data_ptr : data_ptr + this_block_size]
                
                # 테두리 및 서식 적용
                d_height = float(self.config.get('ROW_HEIGHT_DATA', 20.55))
                for r_offset in range(actual_block_rows):
                    r = current_row + r_offset
                    rd = ws.row_dimensions[r]; rd.height = d_height
                    for c in range(1, 14):
                        cell = ws.cell(row=r, column=c)
                        l_s = thin_side; r_s = thin_side
                        t_s = medium_side if r == self.config['START_ROW'] else thin_side
                        b_s = medium_side if r == self.config['DATA_END_ROW'] else thin_side
                        
                        if c == 1: l_s = medium_side
                        if c == 13: r_s = medium_side
                        
                        if c <= 5: # A-E 수직 병합 구역 내부 선 제거
                            if 1 < c < 5: r_s = Side(style=None); l_s = Side(style=None)
                            elif c == 1: r_s = Side(style=None)
                            elif c == 5: l_s = Side(style=None)
                            
                            # 가변 블록 크기에 따른 병합 구역 내부 선 처리
                            if actual_block_rows > 1:
                                if r_offset == 0: b_s = Side(style=None)
                                elif r_offset == actual_block_rows - 1: t_s = Side(style=None)
                                else: t_s = Side(style=None); b_s = Side(style=None)
                                
                        cell.border = Border(left=l_s, right=r_s, top=t_s, bottom=b_s)

                # A-E 수평/수직 병합 및 데이터 입력 (블록 전체 범위)
                # [FIX] actual_block_rows가 1이더라도 A~E는 항상 수평 병합되어야 함
                self.safe_merge_cells(ws, start_row=current_row, start_column=1, end_row=current_row + actual_block_rows - 1, end_column=5)
                
                # F열(6)은 2행 이상일 때만 수직 병합
                if actual_block_rows > 1:
                    self.safe_merge_cells(ws, start_row=current_row, start_column=6, end_row=current_row + actual_block_rows - 1, end_column=6)
                
                # 도면번호/조인트번호 입력 (배치의 첫번째 데이터 기준)
                if batch:
                    # [NEW] Dynamic Column Mapping for PMI (Block level)
                    dwg_col = self.col_to_num(self.config.get('PMI_COL_DWG', '1'))
                    joint_col = self.col_to_num(self.config.get('PMI_COL_JOINT', '6'))
                    
                    if dwg_col >= 1:
                        self.safe_set_value(ws, ws.cell(row=current_row, column=dwg_col).coordinate, batch[0].get('Dwg', ''))
                        ws.cell(row=current_row, column=dwg_col).alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                    
                    if joint_col >= 1:
                        self.safe_set_value(ws, ws.cell(row=current_row, column=joint_col).coordinate, batch[0].get('Joint', batch[0].get('No', '')))
                        ws.cell(row=current_row, column=joint_col).alignment = Alignment(horizontal='center', vertical='center')

                # 개별 데이터 기입
                for i, item in enumerate(batch):
                    r = current_row + i
                    loc_col = self.col_to_num(self.config.get('PMI_COL_LOC', '7'))
                    if loc_col >= 1:
                        self.safe_set_value(ws, ws.cell(row=r, column=loc_col).coordinate, item.get('Loc', ''), align='center')
                    
                    for val_key, col_idx in print_elements:
                        if col_idx < 1: continue
                        raw_v = item.get(val_key, 0.0)
                        v = self.to_float(raw_v)
                        cell = ws.cell(row=r, column=col_idx)
                        self.safe_set_value(ws, cell.coordinate, v if v > 0 else "", align='center')
                        if v > 0: cell.number_format = '0.00'
                    
                    res_col = self.col_to_num(self.config.get('PMI_COL_RES', '13'))
                    if res_col >= 1:
                        self.safe_set_value(ws, ws.cell(row=r, column=res_col).coordinate, item.get('Grade', ''), align='center')
                        cell_l = ws.cell(row=r, column=res_col); cell_l.font = Font(size=8.5); dv_q.add(cell_l)

                data_ptr += len(batch)
                current_row += actual_block_rows
                self.progress['value'] = 30 + (data_ptr / len(all_extracted_data)) * 65

            # [NEW] 마지막 데이터 기입 직후 다음 행 H~L에 BLANK 표시 (병합 적용)
            try:
                # current_row는 이미 다음 3행 블록의 시작점이므로, 실제 마지막 데이터 r의 다음 행을 계산
                last_data_r = (current_row - actual_block_rows) + len(batch)
                if last_data_r <= self.config['DATA_END_ROW']:
                    # H(8)~L(12) 열 병합 후 BLANK 입력
                    self.safe_merge_cells(ws, start_row=last_data_r, start_column=8, end_row=last_data_r, end_column=12)
                    self.safe_set_value(ws, ws.cell(row=last_data_r, column=8).coordinate, "BLANK")
                    ws.cell(row=last_data_r, column=8).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=last_data_r, column=8).font = Font(size=9, bold=True)
            except: pass

            # [RE-NC] 데이터가 없는 빈 칸도 45열까지 3개 행 블록 단위로 테두리/병합 적용
            d_height = float(self.config.get('ROW_HEIGHT_DATA', 20.55))
            while current_row <= self.config['DATA_END_ROW']:
                rows_left = self.config['DATA_END_ROW'] - current_row + 1
                this_block_size = min(3, rows_left)
                
                for r_offset in range(this_block_size):
                    r = current_row + r_offset
                    rd = ws.row_dimensions[r]; rd.height = d_height
                    for c in range(1, 14):
                        cell = ws.cell(row=r, column=c)
                        l_s = thin_side; r_s = thin_side
                        t_s = medium_side if r == self.config['START_ROW'] else thin_side
                        b_s = medium_side if r == self.config['DATA_END_ROW'] else thin_side
                        
                        if c == 1: l_s = medium_side
                        if c == 13: r_s = medium_side
                        
                        # [FIX] 기입되지 않은 행이라도 도면번호(A~E) 수평 병합은 유지
                        if 1 <= c <= 5:
                            if 1 < c < 5: r_s = Side(style=None); l_s = Side(style=None)
                            elif c == 1: r_s = Side(style=None)
                            elif c == 5: l_s = Side(style=None)
                            
                        cell.border = Border(left=l_s, right=r_s, top=t_s, bottom=b_s)
                    
                    # 매 행마다 A~E 수평 병합 실행
                    self.safe_merge_cells(ws, start_row=r, start_column=1, end_row=r, end_column=5)
                
                dv_q.add(ws.cell(row=current_row, column=13))
                current_row += this_block_size

            # [NEW] 서식 정리 도구
            def clear_borders_in_range(sheet, start_r, end_r, start_c=1, end_c=15):
                for r_idx in range(start_r, end_r + 1):
                    for c_idx in range(start_c, end_c + 1):
                        try: sheet.cell(row=r_idx, column=c_idx).border = Border()
                        except: pass

            # [FORCE] 모든 데이터 시트의 바닥선 및 하단 서식 정리
            data_end_row = int(float(self.config.get('DATA_END_ROW', 45)))
            for idx, s in enumerate(wb.worksheets):
                # 데이터 영역 아래쪽(60행까지)의 불필요한 선 제거
                clear_borders_in_range(s, data_end_row + 1, data_end_row + 15)
                
                if s.max_row >= data_end_row:
                    for c in range(1, 14):
                        cell = s.cell(row=data_end_row, column=c)
                        curr_border = cell.border
                        # [CRITICAL] Force synchronize UI variables to config before generation
                        self.save_settings()
                        l_s = curr_border.left; r_s = curr_border.right; t_s = curr_border.top
                        
                        if idx == 0: # 갑지
                            if c in [1, 2, 3, 11, 12, 13]: # A, B, C, K, L, M 바닥선 제거
                                b_s = Side(style=None)
                            else: # D~J 는 약한선
                                b_s = Side(style='thin')
                        else: # 을지
                            b_s = Side(style='medium')
                            
                        cell.border = Border(left=l_s, right=r_s, top=t_s, bottom=b_s)

            # [FORCE] 을지 시트 상단 외곽 테두리 제거 (1행 Top 및 A1-A4 Left, T1-T4 Right)
            for idx, s in enumerate(wb.worksheets):
                if idx > 0: # 을지 시트
                    # 1행 상단 테두리 제거 (T열까지 확장)
                    for c_idx in range(1, 21):
                        cell = s.cell(row=1, column=c_idx)
                        b = cell.border
                        cell.border = Border(left=b.left, right=b.right, top=Side(style=None), bottom=b.bottom)
                    
                    # A1-A4 좌측, M1-M4 우측, T1-T4 우측 테두리 제거
                    for r_idx in range(1, 5):
                        # A열 (좌측)
                        cell_a = s.cell(row=r_idx, column=1)
                        b_a = cell_a.border
                        cell_a.border = Border(left=Side(style=None), right=b_a.right, top=b_a.top, bottom=b_a.bottom)
                        
                        # M열 (우측)
                        cell_m = s.cell(row=r_idx, column=13)
                        b_m = cell_m.border
                        cell_m.border = Border(left=b_m.left, right=Side(style=None), top=b_m.top, bottom=b_m.bottom)
                        
                        # T열 (우측)
                        cell_t = s.cell(row=r_idx, column=20)
                        b_t = cell_t.border
                        cell_t.border = Border(left=b_t.left, right=Side(style=None), top=b_t.top, bottom=b_t.bottom)
                    
                    # [NEW] L3, M3 (페이지 번호 구역) 모든 테두리 제거
                    for c_idx in [12, 13]: # L, M
                        s.cell(row=3, column=c_idx).border = Border()
                    
                    # [NEW] 5행 위쪽 선 진한선으로 (A~M열 범위)
                    for c_idx in range(1, 14): # A~M
                        cell_5 = s.cell(row=5, column=c_idx)
                        b_5 = cell_5.border
                        cell_5.border = Border(left=b_5.left, right=b_5.right, top=medium_side, bottom=b_5.bottom)

            # [FORCE] 을지 시트의 지정 행 글꼴 크기 통일
            for idx, s in enumerate(wb.worksheets):
                if idx > 0: # 을지 시트
                    for r_idx in range(data_start_row, data_end_row + 1):
                        for c_idx in range(1, 14):
                            cell = s.cell(row=r_idx, column=c_idx)
                            f = cell.font
                            if f:
                                cell.font = Font(name=f.name, size=10, bold=f.bold, italic=f.italic, vertAlign=f.vertAlign, underline=f.underline, strike=f.strike, color=f.color)
                            else:
                                cell.font = Font(name='맑은 고딕', size=10)

            total_p = len(wb.worksheets)
            for p_idx, s in enumerate(wb.worksheets):
                page_num = p_idx + 1
                # [NEW] 최종 저장 직전 사용자의 행/열 커스텀 설정을 모든 시트에 다시 한번 적용 (최우선순위)
                ctx = "COVER" if p_idx == 0 else "DATA"
                self.apply_custom_dimensions(s, ctx)
                
                try:
                    if p_idx == 0: # Gapji (Cover)
                        self.safe_merge_cells(s, 3, 14, 3, 20) # N3:T3
                        cell = s["N3"]
                    else: # Eulji (Data)
                        # [FIX] 을지 시트 페이지 번호 가독성 확보를 위해 L3:M3 병합 검토 (또는 M3 단독)
                        try: self.safe_merge_cells(s, 3, 12, 3, 13) # L3:M3
                        except: pass
                        cell = s["L3"] if "L3" in s.merged_cells else s["M3"]
                    
                    cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                    cell.font = Font(name='맑은 고딕', size=11, bold=True)
                    # 단어 사이의 간격을 인위적으로 늘림
                    self.safe_set_value(s, cell.coordinate, f"Page    {page_num}    of    {total_p}")
                except Exception as e:
                    self.log(f"페이지 번호 기입 실패 ({p_idx}): {e}")

            now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_name = f"{os.path.splitext(os.path.basename(template_path))[0]}_Unified_{now_str}{os.path.splitext(template_path)[1]}"
            save_path = os.path.join(os.path.dirname(template_path), output_name); wb.save(save_path)
            self.progress['value'] = 100; self.log(f"✨ 완료! 저장됨: {output_name}")
            messagebox.showinfo("성공", f"통합 성적서 생성이 완료되었습니다.\n\n경로: {os.path.dirname(save_path)}\n파일명: {output_name}")
        except Exception as e:
            self.log(f"❌ 오류: {e}")
            traceback.print_exc()
            # [NEW] 사용자에게 친절한 팝업 알림 추가
            error_msg = f"성적서 생성 중 오류가 발생했습니다.\n\n원인: {e}\n\n* 주로 양식 파일의 시트 구조나 병합 상태가 맞지 않을 때 발생합니다. 다른 양식을 시도하거나 설정을 확인해주세요."
            messagebox.showerror("생성 실패", error_msg)
            # End of _run_pmi_process logic
        finally:
            if 'wb' in locals() and wb:
                try:
                    if hasattr(wb, 'vba_archive') and wb.vba_archive:
                        wb.vba_archive.close()
                except: pass
                try: wb.close()
                except: pass
            for f in glob.glob(os.path.join(tempfile.gettempdir(), "temp_*.png")):
                try: os.remove(f)
                except: pass

    def _run_pt_process(self, final_list, template_path):
        """PT 성적서 생성 (1-row-per-data 레이아웃 + ISO 병합)"""
        self.save_settings() # Ensure UI -> config sync
        self.log(f"🚀 PT 성적서 생성 시작 (총 {len(final_list)} 건)...")
        self.progress['value'] = 0
        try:
            wb = openpyxl.load_workbook(template_path)
            if len(wb.worksheets) < 1:
                raise ValueError("선택한 템플릿 파일에 시트가 존재하지 않습니다.")

            # 갑지 (Cover)
            ws0 = wb.worksheets[0]
            self.add_logos_to_sheet(ws0, is_cover=True, clear_existing=False, mode="PT")
            self._write_gapji_metadata(ws0)
            self.force_print_settings(ws0, context="COVER")
            self.apply_custom_dimensions(ws0, "COVER")

            # 을지 (Data)
            data_sheet_id = 1 if len(wb.worksheets) >= 2 else 0
            ws = wb.worksheets[data_sheet_id]
            ws.title = f"{ws.title[:20]}_001"
            # [FIX] Do NOT clear existing images if data sheet is the same as cover
            self.add_logos_to_sheet(ws, is_cover=False, clear_existing=(ws != ws0), mode="PT")
            self.force_print_settings(ws, context="DATA")

            # 헤더 감지 및 시작 행 결정
            start_row = int(self.config.get('PT_START_ROW', 18))
            end_row = int(self.config.get('PT_END_ROW', 37))
            
            # 스타일 설정
            thin_side = Side(style='thin')
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            # [Smart Detection] 템플릿에서 헤더를 찾아 시작 행 자동 보정
            detected_start = -1
            iso_col_idx = 2 # 기본값 B컬럼
            for r in range(1, 40):
                line_cells = [str(ws.cell(row=r, column=c).value).upper() for c in range(1, 15)]
                line_str = "".join(line_cells)
                if (("JOINT" in line_str or "WELD" in line_str) and ("ISO" in line_str or "LINE" in line_str)) or ("DWG" in line_str and "RESULT" in line_str):
                    detected_start = r + 1
                    # ISO 컬럼 위치 동적 확인
                    for c_idx, val in enumerate(line_cells):
                        if any(k in val for k in ["ISO", "LINE", "DWG", "DRAWING"]) and not any(k in val for k in ["JOINT", "WELD"]):
                            iso_col_idx = c_idx + 1
                            break
                    break
            if detected_start > 0: start_row = detected_start

            current_row = start_row
            current_page = 1
            data_ptr = 0
            
            while data_ptr < len(final_list):
                if current_row > end_row:
                    current_page += 1
                    ws = self.prepare_next_sheet(wb, data_sheet_id, current_page)
                    current_row = start_row
                
                item = final_list[data_ptr]
                
                # 데이터 기입 (B, C, D 컬럼 병합 대응)
                no_col = self.col_to_num(self.config.get('PT_COL_NO', '1'))
                if no_col >= 1:
                    self.safe_set_value(ws, ws.cell(row=current_row, column=no_col).coordinate, item.get('No', ''))
                
                # ISO Drawing No. (병합된 칸 반영)
                dwg_col = self.col_to_num(self.config.get('PT_COL_DWG', str(iso_col_idx)))
                if dwg_col >= 1:
                    iso_cell = ws.cell(row=current_row, column=dwg_col)
                    self.safe_set_value(ws, iso_cell.coordinate, item.get('Dwg', ''))
                    # Preserve original 3-column merge if using default or close to it
                    if dwg_col == iso_col_idx:
                        self.safe_merge_cells(ws, current_row, dwg_col, current_row, dwg_col + 2)
                
                joint_col = self.col_to_num(self.config.get('PT_COL_JOINT', '5'))
                if joint_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=joint_col).coordinate, item.get('Joint', ''))
                
                nps_col = self.col_to_num(self.config.get('PT_COL_NPS', '6'))
                if nps_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=nps_col).coordinate, item.get('NPS', ''))
                
                thk_col = self.col_to_num(self.config.get('PT_COL_THK', '7'))
                if thk_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=thk_col).coordinate, item.get('Thk.', ''))
                
                mat_col = self.col_to_num(self.config.get('PT_COL_MAT', '8'))
                if mat_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=mat_col).coordinate, item.get('Material', ''))
                
                wld_col = self.col_to_num(self.config.get('PT_COL_WELDER', '9'))
                if wld_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=wld_col).coordinate, item.get('Welder', ''))
                
                typ_col = self.col_to_num(self.config.get('PT_COL_TYPE', '10'))
                if typ_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=typ_col).coordinate, item.get('WType', ''))
                
                res_col = self.col_to_num(self.config.get('PT_COL_RES', '11'))
                if res_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=res_col).coordinate, item.get('Result', 'Acc'))
                
                # 스타일링
                for c in range(1, 12):
                    cell = ws.cell(row=current_row, column=c)
                    cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                    cell.font = Font(name='바탕', size=9)
                    cell.border = thin_border

                data_ptr += 1
                current_row += 1
                self.progress['value'] = (data_ptr / len(final_list)) * 95

            total_p = len(wb.worksheets)
            # [NEW] 서식 정리 (48행 유령 선 제거용)
            pt_data_end_row = int(float(self.config.get('PT_END_ROW', 37)))
            for p_idx, s in enumerate(wb.worksheets):
                if p_idx > 0: # 을지
                    for r_idx in range(pt_data_end_row + 1, pt_data_end_row + 15):
                        for c_idx in range(1, 15):
                            try: s.cell(row=r_idx, column=c_idx).border = Border()
                            except: pass
                
                page_num = p_idx + 1
                # 페이지 번호 기입
                try:
                    p_text = f"Page    {page_num}    of    {total_p}"
                    # if p_idx == 0: self.safe_set_value(s, 'O35', p_text)
                    if p_idx > 0: self.safe_set_value(s, 'V3', p_text)
                except: pass

            now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_name = f"PT_Report_{now_str}.xlsx"
            save_path = os.path.join(os.path.dirname(template_path), output_name)
            wb.save(save_path)
            self.progress['value'] = 100
            self.log(f"✨ PT 완료! 저장됨: {output_name}")
            messagebox.showinfo("성공", f"PT 성적서 생성이 완료되었습니다.\n경로: {os.path.dirname(save_path)}")
        except Exception as e:
            self.log(f"❌ PT 생성 오류: {e}")
            traceback.print_exc()
            messagebox.showerror("생성 실패", f"PT 성적서 생성 중 오류가 발생했습니다: {e}")
        finally:
            if 'wb' in locals() and wb:
                try:
                    if hasattr(wb, 'vba_archive') and wb.vba_archive:
                        wb.vba_archive.close()
                except: pass
                try: wb.close()
                except: pass
            for f in glob.glob(os.path.join(tempfile.gettempdir(), "temp_*.png")):
                try: os.remove(f)
                except: pass

    def _run_rt_process(self, final_list, template_path, mode="RT"):
        """RT 성적서 생성 (1-row-per-data 레이아웃)"""
        # [CRITICAL] Force synchronize UI variables to config before generation
        self.save_settings() 
            
        self.log(f"🚀 {mode} 성적서 생성 시작 (총 {len(final_list)} 건)...")
        self.progress['value'] = 0
        try:
            wb = openpyxl.load_workbook(template_path)
            if len(wb.worksheets) < 1:
                raise ValueError("선택한 템플릿 파일에 시트가 존재하지 않습니다.")

            self.log(f"?? [DEBUG] {mode} 갑지 로고 삽입 시도 중...")
            # Gapji (Cover) Logic
            ws0 = wb.worksheets[0]
            self.add_logos_to_sheet(ws0, is_cover=True, clear_existing=False, mode=mode)
            self._write_gapji_metadata(ws0)
            self.force_print_settings(ws0, context=f"{mode}_COVER")
            self.apply_custom_dimensions(ws0, f"{mode}_COVER")

            # Data Sheet Logic (1-row-per-data)
            data_sheet_id = 1 if len(wb.worksheets) >= 2 else 0
            ws = wb.worksheets[data_sheet_id]
            ws.title = f"{ws.title[:20]}_001"
            # [FIX] Do NOT clear existing images if data sheet is the same as cover
            self.add_logos_to_sheet(ws, is_cover=False, clear_existing=(ws != ws0), mode=mode)
            self.force_print_settings(ws, context=f"{mode}_DATA")
            # RT usually doesn't have the same headers as PMI
            # self.set_eulji_headers(ws) 

            # Configurable or Default RT boundaries
            is_kogas = (mode == "KOGAS")
            start_row = int(self.config.get(f'{mode}_START_ROW', 14 if is_kogas else 11))
            end_row = int(self.config.get(f'{mode}_DATA_END_ROW', 25 if is_kogas else 34))
            block_size = 2 if is_kogas else 1
            
            current_row = start_row
            current_page = 1
            data_ptr = 0
            
            while data_ptr < len(final_list):
                # 페이지 넘김 체크: 가스공사는 2행 블록 기준
                if current_row + block_size - 1 > end_row:
                    current_page += 1
                    ws = self.prepare_next_sheet(wb, data_sheet_id, current_page)
                    current_row = start_row
                
                item = final_list[data_ptr]
                
                # [DISABLED] Do NOT write headers automatically to preserve template's own headers
                # if current_row == start_row:
                #     try:
                #         h_row = start_row - 1
                #         if h_row >= 1:
                #             h_map = {
                #                 'RT_COL_NO': 'RT_NAME_NO', 'RT_COL_DATE': 'RT_NAME_DATE',
                #                 'RT_COL_DWG': 'RT_NAME_DWG', 'RT_COL_JOINT': 'RT_NAME_JOINT',
                #                 'RT_COL_LOC': 'RT_NAME_LOC', 'RT_COL_THK': 'RT_NAME_THK',
                #                 'RT_COL_MAT': 'RT_NAME_MAT', 'RT_COL_ACC': 'RT_NAME_ACC',
                #                 'RT_COL_REJ': 'RT_NAME_REJ', 'RT_COL_DEG': 'RT_NAME_DEG',
                #                 'RT_COL_RES': 'RT_NAME_RES', 'RT_COL_WELDER': 'RT_NAME_WELDER',
                #                 'RT_COL_REM': 'RT_NAME_REM'
                #             }
                #             for c_key, n_key in h_map.items():
                #                 c_idx = self.col_to_num(self.config.get(c_key, '0'))
                #                 if c_idx >= 1:
                #                     h_default = c_key.replace('RT_COL_', '').capitalize()
                #                     h_text = self.config.get(n_key, h_default)
                #                     self.safe_set_value(ws, ws.cell(row=h_row, column=c_idx).coordinate, h_text)
                #             for d_i in range(1, 16):
                #                 d_c_idx = self.col_to_num(self.config.get(f'RT_COL_D{d_i}', '0'))
                #                 if d_c_idx >= 1:
                #                     d_text = self.config.get(f'RT_NAME_D{d_i}', f'D{d_i}')
                #                     self.safe_set_value(ws, ws.cell(row=h_row, column=d_c_idx).coordinate, d_text)
                #     except: pass

                if is_kogas:
                    # ===== 가스공사 전용: 2행 1세트 블록 주입 (KOGAS_W_COL_... 적용) =====
                    row_top = current_row      # 상단 행 (Pipe 1 정보)
                    row_bot = current_row + 1  # 하단 행 (Pipe 2 정보)

                    # 순번(No) - 2행 수직 병합
                    no_col = self.col_to_num(self.config.get('KOGAS_W_COL_NO', '1'))
                    if no_col >= 1:
                        try: self.safe_merge_cells(ws, row_top, no_col, row_bot, no_col)
                        except: pass
                        self.safe_set_value(ws, ws.cell(row=row_top, column=no_col).coordinate, item.get('No', ''))
                        ws.cell(row=row_top, column=no_col).alignment = Alignment(horizontal='center', vertical='center')

                    # 용접부(Joint) - 2행 수직 병합
                    joint_col = self.col_to_num(self.config.get('KOGAS_W_COL_JOINT', '4'))
                    if joint_col >= 1:
                        try: self.safe_merge_cells(ws, row_top, joint_col, row_bot, joint_col)
                        except: pass
                        self.safe_set_value(ws, ws.cell(row=row_top, column=joint_col).coordinate, item.get('Joint', ''))
                        ws.cell(row=row_top, column=joint_col).alignment = Alignment(horizontal='center', vertical='center')

                    # 검사일(Date) - 2행 수직 병합
                    date_col = self.col_to_num(self.config.get('KOGAS_W_COL_DATE', '2'))
                    if date_col >= 1:
                        try: self.safe_merge_cells(ws, row_top, date_col, row_bot, date_col)
                        except: pass
                        self.safe_set_value(ws, ws.cell(row=row_top, column=date_col).coordinate, item.get('Date', ''))
                        ws.cell(row=row_top, column=date_col).alignment = Alignment(horizontal='center', vertical='center')

                    # 도면번호(Dwg) - 상단: 주 파이프, 하단: 보조(_Sub)
                    dwg_col = self.col_to_num(self.config.get('KOGAS_W_COL_DWG', '3'))
                    if dwg_col >= 1:
                        self.safe_set_value(ws, ws.cell(row=row_top, column=dwg_col).coordinate, item.get('Dwg', ''))
                        self.safe_set_value(ws, ws.cell(row=row_bot, column=dwg_col).coordinate, item.get('Dwg_Sub', ''))

                    # 재질(Mat) - 상/하단 분리
                    mat_col = self.col_to_num(self.config.get('KOGAS_W_COL_MAT', '7'))
                    if mat_col >= 1:
                        self.safe_set_value(ws, ws.cell(row=row_top, column=mat_col).coordinate, item.get('Mat', ''))
                        self.safe_set_value(ws, ws.cell(row=row_bot, column=mat_col).coordinate, item.get('Mat_Sub', ''))

                    # 용접사(Welder) - 상/하단 분리
                    wld_col = self.col_to_num(self.config.get('KOGAS_W_COL_WELDER', '29'))
                    if wld_col >= 1:
                        self.safe_set_value(ws, ws.cell(row=row_top, column=wld_col).coordinate, item.get('Welder', ''))
                        self.safe_set_value(ws, ws.cell(row=row_bot, column=wld_col).coordinate, item.get('Welder_Sub', ''))

                    # 두께(T) - 2행 수직 병합
                    thk_col = self.col_to_num(self.config.get('KOGAS_W_COL_THK', '6'))
                    if thk_col >= 1:
                        try: self.safe_merge_cells(ws, row_top, thk_col, row_bot, thk_col)
                        except: pass
                        self.safe_set_value(ws, ws.cell(row=row_top, column=thk_col).coordinate, item.get('T', ''))
                        ws.cell(row=row_top, column=thk_col).alignment = Alignment(horizontal='center', vertical='center')

                    # D1~D15 결과 - 상단 행에 개별 지정 열 또는 기본 순차 열 기입
                    for d_i in range(1, 16):
                        d_col_idx = self.col_to_num(self.config.get(f'KOGAS_W_COL_D{d_i}', '0'))
                        if d_col_idx == 0:
                            kogas_d_start_col = int(self.config.get('KOGAS_W_D_START_COL', self.config.get('RT_KOGAS_D_START_COL', '17')))
                            d_col_idx = kogas_d_start_col + (d_i - 1)
                        
                        if d_col_idx >= 1:
                            d_val = item.get(f'D{d_i}', '')
                            if d_val:
                                self.safe_set_value(ws, ws.cell(row=row_top, column=d_col_idx).coordinate, d_val)

                    # 합/불(Result) - 2행 수직 병합
                    res_col = self.col_to_num(self.config.get('KOGAS_W_COL_RES', '28'))
                    if res_col >= 1:
                        try: self.safe_merge_cells(ws, row_top, res_col, row_bot, res_col)
                        except: pass
                        self.safe_set_value(ws, ws.cell(row=row_top, column=res_col).coordinate, item.get('Result', 'ACC'))
                        ws.cell(row=row_top, column=res_col).alignment = Alignment(horizontal='center', vertical='center')

                    # 비고(Remarks) - 2행 수직 병합
                    rem_col = self.col_to_num(self.config.get('KOGAS_W_COL_REM', '30'))
                    if rem_col >= 1:
                        try: self.safe_merge_cells(ws, row_top, rem_col, row_bot, rem_col)
                        except: pass
                        self.safe_set_value(ws, ws.cell(row=row_top, column=rem_col).coordinate, item.get('Remarks', ''))
                        ws.cell(row=row_top, column=rem_col).alignment = Alignment(horizontal='center', vertical='center')

                    current_row += block_size  # 2행씩 전진

                else:
                    # ===== 일반 모드: 1행 1데이터 =====
                    no_col = self.col_to_num(self.config.get('RT_COL_NO', '1'))
                    if no_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=no_col).coordinate, item.get('No', ''))
                    
                    date_col = self.col_to_num(self.config.get('RT_COL_DATE', '2'))
                    if date_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=date_col).coordinate, item.get('Date', ''))
                    
                    dwg_col = self.col_to_num(self.config.get('RT_COL_DWG', '3'))
                    if dwg_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=dwg_col).coordinate, item.get('Dwg', ''))
                    
                    joint_col = self.col_to_num(self.config.get('RT_COL_JOINT', '4'))
                    if joint_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=joint_col).coordinate, item.get('Joint', ''))
                    
                    loc_col = self.col_to_num(self.config.get('RT_COL_LOC', '5'))
                    if loc_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=loc_col).coordinate, item.get('Loc', ''))
                    
                    thk_col = self.col_to_num(self.config.get('RT_COL_THK', '6'))
                    if thk_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=thk_col).coordinate, item.get('T', ''))
                    
                    mat_col = self.col_to_num(self.config.get('RT_COL_MAT', '7'))
                    if mat_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=mat_col).coordinate, item.get('Mat', ''))
                    
                    acc_col = self.col_to_num(self.config.get('RT_COL_ACC', '7'))
                    if acc_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=acc_col).coordinate, item.get('Acc', ''))
                    
                    rej_col = self.col_to_num(self.config.get('RT_COL_REJ', '8'))
                    if rej_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=rej_col).coordinate, item.get('Rej', ''))
                    
                    deg_col = self.col_to_num(self.config.get('RT_COL_DEG', '9'))
                    if deg_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=deg_col).coordinate, item.get('Deg', ''))

                    iqi_col = self.col_to_num(self.config.get('RT_COL_IQI', '10'))
                    if iqi_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=iqi_col).coordinate, item.get('IQI', ''))
                    
                    sens_col = self.col_to_num(self.config.get('RT_COL_SENS', '11'))
                    if sens_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=sens_col).coordinate, item.get('Sens', ''))
                    
                    den_col = self.col_to_num(self.config.get('RT_COL_DEN', '12'))
                    if den_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=den_col).coordinate, item.get('Den', ''))
                    
                    for d_i in range(1, 16):
                        d_col_key = f'RT_COL_D{d_i}'
                        d_col_idx = self.col_to_num(self.config.get(d_col_key, str(12 + d_i)))
                        if d_col_idx >= 1:
                            self.safe_set_value(ws, ws.cell(row=current_row, column=d_col_idx).coordinate, item.get(f'D{d_i}', ''))
                    
                    res_col = self.col_to_num(self.config.get('RT_COL_RES', '28'))
                    if res_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=res_col).coordinate, item.get('Result', 'ACC'))
                    
                    wld_col = self.col_to_num(self.config.get('RT_COL_WELDER', '29'))
                    if wld_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=wld_col).coordinate, item.get('Welder', ''))
                    
                    rem_col = self.col_to_num(self.config.get('RT_COL_REM', '30'))
                    if rem_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=rem_col).coordinate, item.get('Remarks', ''))

                    current_row += 1

                data_ptr += 1
                self.progress['value'] = (data_ptr / len(final_list)) * 95

            total_p = len(wb.worksheets)
            # [NEW] 서식 정리 [DISABLED] - 템플릿 선(Border) 보존을 위해 주석 처리
            # rt_data_end_row = int(float(self.config.get('RT_DATA_END_ROW', 31)))
            for p_idx, s in enumerate(wb.worksheets):
                # if p_idx > 0: # 을지
                #     for r_idx in range(rt_data_end_row + 1, rt_data_end_row + 15):
                #         for c_idx in range(1, 31):
                #             try: s.cell(row=r_idx, column=c_idx).border = Border()
                #             except: pass
                
                page_num = p_idx + 1
                self.apply_custom_dimensions(s, "DATA" if p_idx > 0 else "COVER")
                # [FIX] RT Page Numbers: Gab-sheet -> N3, Eul-sheet -> S2
                try:
                    p_text = f"Page    {page_num}    of    {total_p}"
                    if p_idx == 0: # Gapji (Cover)
                        self.safe_set_value(s, "N3", p_text)
                        s["N3"].alignment = Alignment(horizontal='center', vertical='center')
                        s["N3"].font = Font(name='맑은 고딕', size=11, bold=True)
                    else: # Eulji (Data)
                        self.safe_set_value(s, "S2", p_text)
                        s["S2"].alignment = Alignment(horizontal='center', vertical='center')
                        s["S2"].font = Font(name='맑은 고딕', size=11, bold=True)
                except Exception as pe:
                    self.log(f"   ⚠️ 페이지 번호 기입 실패 ({p_idx}): {pe}")

            # [FINAL CLEANUP] Standardized cleanup for RT
            try:
                ws0 = wb.worksheets[0]
                sig_end = 48
                # [SAFE] Only clear far below the Shooting Sketch (Rows 65+)
                for r in range(65, 80):
                    # Do NOT unmerge on Gapji anymore to be safe
                    # merged_ranges = [str(rng) for rng in ws0.merged_cells.ranges if rng.min_row >= r]
                    for c in range(1, 21):
                        try: ws0.cell(row=r, column=c).border = Border()
                        except: pass
                self.log("   ✅ RT 갑지 하단 잔여 서식 정리 완료 (65-80행)")
            except: pass

            now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_name = f"RT_Report_{now_str}.xlsx"
            save_path = os.path.join(os.path.dirname(template_path), output_name)
            wb.save(save_path)
            
            # [SURGERY] openpyxl destroys complex drawings (like grouped shapes). 
            # We surgically inject the template's drawing layer back into the generated file.
            self.inject_drawing_layer(template_path, save_path)
            
            self.progress['value'] = 100
            self.log(f"✨ RT 완료! 저장됨: {output_name}")
            messagebox.showinfo("성공", f"RT 성적서 생성이 완료되었습니다.\n경로: {os.path.dirname(save_path)}")
        except Exception as e:
            self.log(f"❌ RT 생성 오류: {e}")
            traceback.print_exc()
            messagebox.showerror("생성 실패", f"RT 성적서 생성 중 오류가 발생했습니다: {e}")
        finally:
            if 'wb' in locals() and wb:
                try:
                    if hasattr(wb, 'vba_archive') and wb.vba_archive:
                        wb.vba_archive.close()
                except: pass
                try: wb.close()
                except: pass
            for f in glob.glob(os.path.join(tempfile.gettempdir(), "temp_*.png")):
                try: os.remove(f)
                except: pass

    # --- PHOTO LOG UI & LOGIC ---
    def _setup_photo_log_ui(self, parent):
        """Standardized Photo Log UI with Dual-Pane Layout."""
        # [SYNC] Match PAUT PanedWindow properties exactly
        self.photo_paned = tk.PanedWindow(parent, orient='horizontal', background="#d1d5db", 
                            sashwidth=6, sashpad=0, sashrelief='raised', borderwidth=0)
        self.photo_paned.pack(fill='both', expand=True)

        # [LEFT] Settings Sidebar
        left_container = tk.Frame(self.photo_paned, background="#f9fafb", highlightthickness=0, borderwidth=0)
        # Add with minwidth to prevent content from being completely hidden
        self.photo_paned.add(left_container, width=425, minsize=200)
        
        # Scrollable area (Full size)
        left_pane = self._create_scrollable_sidebar(left_container)

        # FIXED FLOATING Header
        header_frame = tk.Frame(left_container, background="#f9fafb", highlightthickness=0, borderwidth=0)
        header_frame.place(x=0, y=0, relwidth=1, height=40)
        
        inner_header = tk.Frame(header_frame, background="#f9fafb", padx=20)
        inner_header.pack(fill='both', expand=True, pady=(5, 0))
        
        tk.Label(inner_header, text="📸 사진대장 관리", font=("Malgun Gothic", 15, "bold"), 
                 background="#f9fafb", foreground="#1e3a8a").pack(side='left')
        
        tk.Frame(header_frame, height=1, background="#e5e7eb").pack(side='bottom', fill='x')
        header_frame.lift()

        # Spacer in scrollable area
        tk.Frame(left_pane, height=40, background="#f9fafb").pack(fill='x')

        # 1. Report Info Group
        info_frame = ttk.LabelFrame(left_pane, text=" 리포트 정보 (Report Info) ", padding=10)
        info_frame.pack(fill='x', padx=10, pady=5)

        tk.Label(info_frame, text="검사 항목:").grid(row=0, column=0, sticky='w', pady=2)
        type_combo = ttk.Combobox(info_frame, textvariable=self.photo_inspect_type, 
                                 values=list(self.photo_header_map.keys()), state="readonly")
        type_combo.grid(row=0, column=1, sticky='ew', padx=5, pady=2)
        type_combo.bind("<<ComboboxSelected>>", self._on_photo_type_change)

        tk.Label(info_frame, text="리포트 제목:").grid(row=1, column=0, sticky='w', pady=2)
        ttk.Entry(info_frame, textvariable=self.photo_report_title, width=1).grid(row=1, column=1, sticky='ew', padx=5, pady=2)

        tk.Label(info_frame, text="발주처:").grid(row=2, column=0, sticky='w', pady=2)
        ttk.Entry(info_frame, textvariable=self.photo_orderer, width=1).grid(row=2, column=1, sticky='ew', padx=5, pady=2)

        tk.Label(info_frame, text="리포트 번호:").grid(row=3, column=0, sticky='w', pady=2)
        ttk.Entry(info_frame, textvariable=self.photo_report_no, width=1).grid(row=3, column=1, sticky='ew', padx=5, pady=2)

        tk.Label(info_frame, text="검사 일자:").grid(row=4, column=0, sticky='w', pady=2)
        ttk.Entry(info_frame, textvariable=self.photo_inspect_date, width=1).grid(row=4, column=1, sticky='ew', padx=5, pady=2)
        
        tk.Label(info_frame, text="로고 파일:").grid(row=5, column=0, sticky='w', pady=2)
        logo_f = tk.Frame(info_frame)
        logo_f.grid(row=5, column=1, sticky='ew', padx=5, pady=2)
        ttk.Entry(logo_f, textvariable=self.photo_logo_path).pack(side='left', fill='x', expand=True)
        def browse_logo():
            f = filedialog.askopenfilename(filetypes=[("Images", "*.png;*.jpg;*.jpeg;*.bmp")])
            if f: self.photo_logo_path.set(f)
        ttk.Button(logo_f, text="...", width=3, command=browse_logo).pack(side='right')
        
        info_frame.columnconfigure(1, weight=1)

        # 2. Layout Settings Group
        layout_frame = ttk.LabelFrame(left_pane, text=" 사진 레이아웃 설정 (Layout) ", padding=10)
        layout_frame.pack(fill='x', padx=10, pady=5)

        tk.Label(layout_frame, text="한 줄당 사진:").grid(row=0, column=0, sticky='w', pady=2)
        ttk.Combobox(layout_frame, textvariable=self.photo_cols_per_row, values=["1", "2", "3"], state="readonly", width=5).grid(row=0, column=1, sticky='w', padx=5, pady=2)
        ttk.Checkbutton(layout_frame, text="비율 유지", variable=self.photo_keep_aspect).grid(row=0, column=2, sticky='w')

        tk.Label(layout_frame, text="칸 너비/높이:").grid(row=1, column=0, sticky='w', pady=2)
        wh_f = tk.Frame(layout_frame)
        wh_f.grid(row=1, column=1, columnspan=2, sticky='w')
        ttk.Entry(wh_f, textvariable=self.photo_cell_width_var, width=6).pack(side='left', padx=2)
        ttk.Entry(wh_f, textvariable=self.photo_cell_height_var, width=6).pack(side='left', padx=2)

        tk.Label(layout_frame, text="여백(T/B/L/R):").grid(row=2, column=0, sticky='w', pady=2)
        m_f = tk.Frame(layout_frame)
        m_f.grid(row=2, column=1, columnspan=2, sticky='w')
        ttk.Entry(m_f, textvariable=self.photo_margin_top_var, width=4).pack(side='left', padx=1)
        ttk.Entry(m_f, textvariable=self.photo_margin_bottom_var, width=4).pack(side='left', padx=1)
        ttk.Entry(m_f, textvariable=self.photo_margin_left_var, width=4).pack(side='left', padx=1)
        ttk.Entry(m_f, textvariable=self.photo_margin_right_var, width=4).pack(side='left', padx=1)

        tk.Label(layout_frame, text="설명 높이:").grid(row=3, column=0, sticky='w', pady=2)
        ttk.Entry(layout_frame, textvariable=self.photo_desc_height_var, width=10).grid(row=3, column=1, sticky='w', padx=5, pady=2)
        
        tk.Label(layout_frame, text="인쇄 배율:").grid(row=4, column=0, sticky='w', pady=2)
        ttk.Entry(layout_frame, textvariable=self.photo_print_scale_var, width=10).grid(row=4, column=1, sticky='w', padx=5, pady=2)

        tk.Label(layout_frame, text="배치 설정:").grid(row=5, column=0, sticky='w', pady=2)
        b_f = tk.Frame(layout_frame)
        b_f.grid(row=5, column=1, columnspan=2, sticky='w')
        ttk.Combobox(b_f, textvariable=self.photo_align_var, values=["좌측 정렬", "중앙 정렬"], state="readonly", width=10).pack(side='left', padx=2)
        ttk.Checkbutton(b_f, text="가로 폭 맞춤 (Fit to Width)", variable=self.photo_fit_width_var).pack(side='left', padx=5)
        ttk.Checkbutton(b_f, text="세로 사진 자동 회전", variable=self.photo_auto_rotate_var).pack(side='left', padx=5)
        
        tk.Label(layout_frame, text="너비비율(%):").grid(row=6, column=0, sticky='w')
        wf_f = tk.Frame(layout_frame)
        wf_f.grid(row=6, column=1, columnspan=2, sticky='w')
        ttk.Entry(wf_f, textvariable=self.photo_width_pct_var, width=7).pack(side='left', padx=2)
        tk.Label(wf_f, text="너비추가(px):").pack(side='left', padx=(10, 0))
        ttk.Entry(wf_f, textvariable=self.photo_width_pixel_adj_var, width=4).pack(side='left', padx=2)
        tk.Label(wf_f, text="좌우(px):").pack(side='left', padx=(10, 0))
        ttk.Entry(wf_f, textvariable=self.photo_shift_x_var, width=4).pack(side='left', padx=2)
        tk.Label(wf_f, text="상하(px):").pack(side='left', padx=(10, 0))
        ttk.Entry(wf_f, textvariable=self.photo_shift_y_var, width=4).pack(side='left', padx=2)
        tk.Label(wf_f, text="(100% 기준 미세조정)", font=('', 9), foreground='gray').pack(side='left', padx=5)

        # 3. Logo Options
        logo_frame = ttk.LabelFrame(left_pane, text=" 로고 및 출력 설정 ", padding=10)
        logo_frame.pack(fill='x', padx=10, pady=5)

        tk.Label(logo_frame, text="로고 너비:").grid(row=0, column=0, sticky='w', pady=2)
        ttk.Entry(logo_frame, textvariable=self.photo_logo_width_var, width=10).grid(row=0, column=1, sticky='w', padx=5, pady=2)
        
        tk.Label(logo_frame, text="로고 X/Y:").grid(row=1, column=0, sticky='w', pady=2)
        xy_f = tk.Frame(logo_frame)
        xy_f.grid(row=1, column=1, sticky='w')
        ttk.Entry(xy_f, textvariable=self.photo_logo_x_var, width=5).pack(side='left', padx=2)
        ttk.Entry(xy_f, textvariable=self.photo_logo_y_var, width=5).pack(side='left', padx=2)
        
        # [NEW] Photo Log Layout Save Button
        btn_f = tk.Frame(logo_frame, background="#ffffff")
        btn_f.grid(row=2, column=0, columnspan=2, sticky='ew', pady=5)
        ttk.Button(btn_f, text="💾 사진대장 레이아웃 설정 저장", command=self.save_photo_log_config).pack(side='left', padx=5)
        tk.Label(btn_f, text="* 로고위치, 여백 등이 photolog_config.json에 저장됩니다.", font=("Malgun Gothic", 8), fg="gray", background="#ffffff").pack(side='left')
        
        # [NEW] Save Photo Log Config Button
        btn_f = tk.Frame(logo_frame, background="#ffffff")
        btn_f.grid(row=2, column=0, columnspan=2, sticky='ew', pady=5)
        ttk.Button(btn_f, text="💾 사진대장 레이아웃 설정 저장", command=self.save_photo_log_config).pack(side='left', padx=5)
        tk.Label(btn_f, text="* 여백, 로고위치, 너비비율 등이 저장됩니다.", font=("Malgun Gothic", 8), fg="gray", background="#ffffff").pack(side='left')
        
        tk.Label(logo_frame, text="출력 파일명:").grid(row=2, column=0, sticky='w', pady=2)
        ttk.Entry(logo_frame, textvariable=self.photo_output_name, width=1).grid(row=2, column=1, sticky='ew', padx=5, pady=2)
        logo_frame.columnconfigure(1, weight=1)

        # Action Buttons at bottom of sidebar
        btn_f = tk.Frame(left_pane, background="#f1f5f9")
        btn_f.pack(fill='x', padx=20, pady=10)
        
        ttk.Button(btn_f, text="🚀 사진대장 리포트 생성", style="Accent.TButton", 
                   command=self.start_photo_generation).pack(fill='x', pady=5)
        ttk.Button(btn_f, text="💾 현재 설정 저장", 
                   command=self.save_settings).pack(fill='x', pady=5)

        # [RIGHT] Preview & File List
        right_container = tk.Frame(self.photo_paned, background="#ffffff")
        self.photo_paned.add(right_container)

        # File List Header
        list_header = tk.Frame(right_container, background="#f8fafc", padx=15, pady=10)
        list_header.pack(fill='x')
        tk.Label(list_header, text="📁 선택된 사진 리스트 (파일 순차 정렬됨)", font=("Malgun Gothic", 10, "bold"), 
                 background="#f8fafc", foreground="#475569").pack(side='left')
        
        # Tools under header
        tool_bar = tk.Frame(right_container, background="#ffffff", padx=10, pady=5)
        tool_bar.pack(fill='x')
        ttk.Button(tool_bar, text="파일 개별 추가", command=self._add_photo_files).pack(side='left', padx=2)
        ttk.Button(tool_bar, text="폴더 전체 추가", command=self._add_photo_folder).pack(side='left', padx=2)
        ttk.Button(tool_bar, text="전체 비우기", command=self._clear_photo_all).pack(side='right', padx=2)
        ttk.Button(tool_bar, text="선택 항목 제거", command=self._remove_photo_selected).pack(side='right', padx=2)

        # Listbox area
        list_frame = tk.Frame(right_container, background="#ffffff")
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.photo_listbox = tk.Listbox(list_frame, font=("Consolas", 9), selectmode="extended", borderwidth=1, relief='flat', highlightthickness=1, highlightcolor="#3b82f6")
        self.photo_listbox.pack(side='left', fill='both', expand=True)
        
        list_vsb = ttk.Scrollbar(list_frame, orient="vertical", command=self.photo_listbox.yview)
        # [SYNC] Sash persistence bindings (Match PAUT)
        self.photo_pane_ratio = self.config.get('PHOTO_SASH_RATIO', 0.45)
        self.photo_paned.bind("<Configure>", lambda e: [self._on_photo_paned_configure(e), self.root.update_idletasks()])
        self.photo_paned.bind("<ButtonRelease-1>", lambda e: self.root.after(10, self._update_photo_ratio))
        
        # [RESTORED] Sync listbox if any files were loaded from settings
        if self.photo_selected_files:
            self.photo_listbox.delete(0, tk.END)
            for f in self.photo_selected_files:
                self.photo_listbox.insert(tk.END, f)
                
        self._apply_sash_ratio("PHOTO")
        self.root.after(500, lambda: self._on_photo_paned_configure(None))

    def _update_photo_ratio(self):
        try:
            total_w = self.photo_paned.winfo_width()
            if total_w > 100:
                current_sash = self.photo_paned.sash_coord(0)[0]
                self.photo_pane_ratio = current_sash / total_w
                self.config['PHOTO_SASH_RATIO'] = self.photo_pane_ratio
                self.save_settings()
        except: pass

    def _on_photo_paned_configure(self, event):
        try:
            total_w = self.photo_paned.winfo_width()
            if total_w > 100:
                if event and event.widget == self.photo_paned:
                    new_pos = int(total_w * self.photo_pane_ratio)
                    self.photo_paned.sash_place(0, new_pos, 0)
        except: pass

    def _on_photo_type_change(self, event=None):
        new_type = self.photo_inspect_type.get()
        if new_type in self.photo_header_map:
            self.photo_report_title.set(self.photo_header_map[new_type])

    def _add_photo_files(self):
        files = filedialog.askopenfilenames(filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp")])
        if files:
            for f in files:
                f_norm = os.path.normpath(f)
                if f_norm not in self.photo_selected_files:
                    self.photo_selected_files.append(f_norm)
                    if hasattr(self, 'photo_listbox'):
                        self.photo_listbox.insert(tk.END, f_norm)
            self.log(f"[PhotoLog] {len(files)}개 파일 추가 시도")

    def _add_photo_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            added_count = 0
            for ext in ('*.png', '*.jpg', '*.jpeg', '*.bmp'):
                for f in glob.glob(os.path.join(folder, ext)):
                    f_norm = os.path.normpath(f)
                    if f_norm not in self.photo_selected_files:
                        self.photo_selected_files.append(f_norm)
                        if hasattr(self, 'photo_listbox'):
                            self.photo_listbox.insert(tk.END, f_norm)
                        added_count += 1
            self.log(f"[PhotoLog] 폴더에서 {added_count}개 파일 추가 완료")

    def _remove_photo_selected(self):
        idxs = list(self.photo_listbox.curselection())
        for i in reversed(idxs):
            path = self.photo_listbox.get(i)
            if path in self.photo_selected_files:
                self.photo_selected_files.remove(path)
            self.photo_listbox.delete(i)

    def _clear_photo_all(self):
        self.photo_selected_files.clear()
        self.photo_listbox.delete(0, tk.END)
        self.log("[PhotoLog] 리스트 초기화 완료")

    def start_photo_generation(self):
        if not self.photo_selected_files:
            messagebox.showwarning("경고", "리포트에 포함할 이미지를 먼저 선택해주세요.")
            return
        self.save_settings() # Auto-save before generation
        threading.Thread(target=self.generate_photo_report, daemon=True).start()

    def generate_photo_report(self):
        try:
            if not self.photo_selected_files: return
            
            image_files = sorted(self.photo_selected_files)
            
            # [REMEMBER] Load last save dir from class variable
            initial_folder = self.last_photo_save_dir if self.last_photo_save_dir and os.path.exists(self.last_photo_save_dir) else os.path.dirname(image_files[0])
            
            default_name = self.photo_output_name.get()
            if not default_name.endswith(".xlsx"):
                default_name += ".xlsx"
                
            # [NEW] Ask for save folder and filename
            output_path = filedialog.asksaveasfilename(
                title="사진대장 저장 위치 선택",
                initialdir=initial_folder,
                initialfile=default_name,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not output_path:
                self.log("[PhotoLog] 작업을 취소했습니다.")
                return

            # [REMEMBER] Update last save directory
            self.last_photo_save_dir = os.path.dirname(output_path)
            self.save_settings() # Silent save

            self.progress["value"] = 0
            self.log("[PhotoLog] 작업을 시작합니다..")
            
            workbook = xlsxwriter.Workbook(output_path)
            worksheet = workbook.add_worksheet()

            # Page Setup
            worksheet.set_paper(9) # A4
            worksheet.set_portrait()
            worksheet.center_horizontally()
            
            try:
                m_t = float(self.photo_margin_top_var.get())
                m_b = float(self.photo_margin_bottom_var.get())
                m_l = float(self.photo_margin_left_var.get())
                m_r = float(self.photo_margin_right_var.get())
                worksheet.set_margins(left=m_l, right=m_r, top=m_t, bottom=m_b)
            except:
                worksheet.set_margins(left=0.4, right=0.4, top=0.5, bottom=0.5)
            
            worksheet.set_footer('&C&P / &N')
            worksheet.repeat_rows(0, 4) 

            # Layout Calculation
            num_cols = int(self.photo_cols_per_row.get())
            photos_per_page = 4 if num_cols == 1 else (8 if num_cols == 2 else 12)
            total_pages = math.ceil(len(self.photo_selected_files) / photos_per_page)
            worksheet.fit_to_pages(1, total_pages)

            # Formats
            title_format = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'shrink': True})
            company_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': 9, 'text_wrap': True})
            center_border = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': 10})
            bold_format = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
            desc_format = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'border': 1, 'font_size': 10, 'shrink': True, 'text_wrap': False, 'indent': 1})

            # Fixed 6-column Grid System
            GRID_COLS = 6
            unit_per_grid = (float(self.photo_cell_width_var.get()) * 2) / GRID_COLS
            worksheet.set_column(0, GRID_COLS - 1, unit_per_grid)

            CELL_ROW_HEIGHT = float(self.photo_cell_height_var.get())
            
            # [REFINED] Percentage-Based Precision Scaling
            try:
                WIDTH_PCT = float(self.photo_width_pct_var.get().strip()) / 100.0
                PIXEL_ADJ = float(self.photo_width_pixel_adj_var.get().strip())
                SHIFT_X = float(self.photo_shift_x_var.get().strip())
                SHIFT_Y = float(self.photo_shift_y_var.get().strip())
            except:
                WIDTH_PCT, PIXEL_ADJ, SHIFT_X, SHIFT_Y = 1.0, 0.0, 0.0, 0.0
                
            # Fixed internal factor for stability
            INTERNAL_FACTOR = 7.142
                
            if num_cols == 1:
                photo_col_spans = [(0, GRID_COLS - 1)]
                CELL_WIDTH_PX = round(((unit_per_grid * INTERNAL_FACTOR + 5) * 6), 1)
            elif num_cols == 2:
                photo_col_spans = [(0, 2), (3, 5)]
                CELL_WIDTH_PX = round(((unit_per_grid * INTERNAL_FACTOR + 5) * 3), 1)
            else: # 3 Columns
                photo_col_spans = [(0, 1), (2, 3), (4, 5)]
                CELL_WIDTH_PX = round(((unit_per_grid * INTERNAL_FACTOR + 5) * 2), 1)
            
            # Target width based on percentage and padding
            SAFE_WIDTH = (CELL_WIDTH_PX - 10) * WIDTH_PCT + PIXEL_ADJ
            
            # Detailed Logging for user calibration
            self.log(f"[PhotoLog] 설정: 너비비율={WIDTH_PCT*100:.1f}%, 보정={PIXEL_ADJ}px")
            self.log(f"[PhotoLog] 이동: 가로={SHIFT_X}px, 세로={SHIFT_Y}px")
            self.log(f"[PhotoLog] 계산결과: 셀너비 {CELL_WIDTH_PX}px -> 사진너비 {SAFE_WIDTH:.1f}px")

            worksheet.set_row(0, 30)
            worksheet.merge_range(0, 0, 0, GRID_COLS-1, self.photo_report_title.get(), title_format)
            
            company_text = "서   울   檢   査   株   式   會   社\nSEOUL INSPECTION & TESTING Co., Ltd.\nTEL : (02) 552-1112   FAX : (02) 2058-0720"
            worksheet.merge_range(1, 0, 3, 2, company_text, company_format)

            # Logo Insertion with Smart Fallback
            logo_f = self.photo_logo_path.get()
            if not logo_f or not os.path.exists(logo_f):
                logo_f = self.find_image_smart("SITCO")
                if not logo_f: logo_f = self.find_image_smart("서울검사")
            
            if logo_f and os.path.exists(logo_f):
                try:
                    total_header_h = 45 # 15 * 3 rows
                    for r in range(1, 4): worksheet.set_row(r, 15) 
                    with PILImage.open(logo_f) as img:
                        w, h = img.size
                        max_w_logo = float(self.photo_logo_width_var.get())
                        mx = float(self.photo_logo_x_var.get())
                        my = float(self.photo_logo_y_var.get())
                        scale = min(max_w_logo/w, 42/h) * 0.95
                        logo_h = h * scale
                        y_offset = (total_header_h - logo_h) / 2 + my
                        worksheet.insert_image('A2', logo_f, {'x_scale': scale, 'y_scale': scale, 'x_offset': mx, 'y_offset': y_offset, 'object_position': 1})
                except Exception as e:
                    self.log(f"[PhotoLog] 로고 삽입 중 오류: {e}")

            worksheet.merge_range(1, 3, 1, 5, f"발주처: {self.photo_orderer.get()}", center_border)
            worksheet.merge_range(2, 3, 2, 5, f"REPORT NO: {self.photo_report_no.get()}", center_border)
            worksheet.merge_range(3, 3, 3, 5, f"검사일자: {self.photo_inspect_date.get()}", center_border)
            
            worksheet.set_row(4, 25)
            worksheet.merge_range(4, 0, 4, GRID_COLS-1, "PHOTO LOG (사진 대장)", bold_format)

            row = 5
            col_ptr = 0
            page_breaks = []
            photos_per_page = 4 if num_cols == 1 else (8 if num_cols == 2 else 12)
            DESC_ROW_HEIGHT = float(self.photo_desc_height_var.get())
            CELL_HEIGHT_PX = (CELL_ROW_HEIGHT * 1.33333) - 2
            
            total = len(image_files)
            # Keep track of the maximum height needed for the current row
            current_row_max_h_pt = CELL_ROW_HEIGHT
            ROW_PT_TO_PX = 1.33333

            for i, img_path in enumerate(image_files):
                # Reset max height for a new row of photos
                if col_ptr == 0:
                    current_row_max_h_pt = CELL_ROW_HEIGHT
                
                try:
                    with PILImage.open(img_path) as img:
                        img = ImageOps.exif_transpose(img)
                        img_w, img_h = img.size
                        
                        # [NEW] Auto-rotate vertical images to horizontal if option enabled
                        if self.photo_auto_rotate_var.get() and img_h > img_w:
                            img = img.rotate(90, expand=True)
                            img_w, img_h = img.size
                            self.log(f"[PhotoLog] 자동 회전 적용: {os.path.basename(img_path)}")
                        
                        # [FIX] Prepare image data to preserve rotation in Excel
                        img_buffer = io.BytesIO()
                        img.save(img_buffer, format='PNG')
                        img_buffer.seek(0)
                        
                        c_start, c_end = photo_col_spans[col_ptr]
                        if c_start != c_end: worksheet.merge_range(row, c_start, row, c_end, "", center_border)
                        
                        # [REFINED] Scaling Logic with 'Fit to Width' support
                        TOTAL_BUFFER = 10 # Aggressive Safety Margin
                        if self.photo_fit_width_var.get():
                            x_scale = SAFE_WIDTH / img_w
                            y_scale = x_scale
                            # Calculate required height for this photo
                            req_h_px = (img_h * y_scale) + 10 # Buffer
                            req_h_pt = req_h_px / ROW_PT_TO_PX
                            current_row_max_h_pt = max(current_row_max_h_pt, req_h_pt)
                        elif not self.photo_keep_aspect.get():
                            x_scale = SAFE_WIDTH / img_w
                            y_scale = (CELL_HEIGHT_PX - 10) / img_h
                        else:
                            scale = min(SAFE_WIDTH / img_w, (CELL_HEIGHT_PX - 10) / img_h)
                            x_scale = y_scale = scale
                        
                        # Apply the potentially updated row height
                        worksheet.set_row(row, current_row_max_h_pt)
                        
                        # Re-calculate Y offset based on the final row height
                        final_row_h_px = current_row_max_h_pt * ROW_PT_TO_PX
                        
                        # [UNIFIED] Uniform Margin Distribution with Manual Shift (High Precision)
                        x_off_float = ((CELL_WIDTH_PX - (img_w * x_scale)) / 2) + SHIFT_X
                        y_off_float = ((final_row_h_px - (img_h * y_scale)) / 2) + SHIFT_Y
                        
                        x_off = round(x_off_float)
                        y_off = round(y_off_float)
                        
                        # Log precision offsets for user calibration
                        self.log(f"[PhotoLog] 사진 배치: {os.path.basename(img_path)}")
                        self.log(f"           - 가로위치: {x_off_float:.2f} -> {x_off}px")
                        self.log(f"           - 세로위치: {y_off_float:.2f} -> {y_off}px")
                        
                        worksheet.insert_image(row, c_start, img_path, {
                            'image_data': img_buffer,
                            'x_scale': x_scale, 'y_scale': y_scale, 
                            'x_offset': x_off, 'y_offset': y_off, 
                            'object_position': 2 # Move but don't size with cells
                        })
                except Exception as e:
                    self.log(f"[Error] {os.path.basename(img_path)}: {e}")

                name = os.path.splitext(os.path.basename(img_path))[0]
                worksheet.set_row(row + 1, DESC_ROW_HEIGHT)
                c_start, c_end = photo_col_spans[col_ptr]
                worksheet.merge_range(row+1, c_start, row+1, c_end, f"사진 설명: {name}", desc_format)
                
                col_ptr += 1
                if col_ptr >= num_cols:
                    col_ptr = 0
                    row += 2
                
                if (i + 1) % photos_per_page == 0 and (i + 1) < total:
                    page_breaks.append(row if col_ptr==0 else row+2)

                self.progress["value"] = ((i + 1) / total) * 100
                self.log(f"[PhotoLog] 처리 중.. ({i+1}/{total})")

            if page_breaks: worksheet.set_h_pagebreaks(page_breaks)
            workbook.close()
            self.log(f"[PhotoLog] 완료: {os.path.basename(output_path)}")
            messagebox.showinfo("성공", f"사진대장이 생성되었습니다.\n{output_path}")

            # [NEW] Open the folder
            try:
                os.startfile(os.path.dirname(output_path))
            except: pass

        except Exception as e:
            self.log(f"[Error] {e}")
            messagebox.showerror("오류", f"작업 중 오류 발생:\n{e}")

    def _on_entry_esc(self, event):
        """Removes focus and clears selection on ESC (preserves text)."""
        try:
            widget = event.widget
            if hasattr(widget, 'selection_clear'):
                widget.selection_clear()
            self.root.focus_set()
        except: pass

if __name__ == "__main__":
    root = tk.Tk()
    app = PMIReportApp(root)
    root.mainloop()
