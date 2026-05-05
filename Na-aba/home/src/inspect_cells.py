import openpyxl
import os

current_dir = r"c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\Na-aba\home\src"
home_dir = os.path.dirname(current_dir)
template_path = os.path.join(home_dir, 'resources', 'Template_DailyWorkReport.xlsx')

wb = openpyxl.load_workbook(template_path, data_only=True)
sheet = wb.active

for r in [29, 30]:
    for c in range(5, 16): # E to O
        col = openpyxl.utils.get_column_letter(c)
        val = sheet.cell(row=r, column=c).value
        print(f"{col}{r}: [{val}]")
