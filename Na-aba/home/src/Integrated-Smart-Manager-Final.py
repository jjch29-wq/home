# -*- coding: utf-8 -*-
### VERSION: INTEGRATED_FINAL_FULL_MERGE_V1 ###
import mimetypes
import os
import sys
import zipfile
import shutil
import subprocess
import time
import datetime
import json
import ctypes
import re
import traceback
import warnings
import glob
import math
import tempfile
import io
from PIL import Image as PILImage, ImageChops, ImageOps, ImageTk
from openpyxl.cell.cell import MergedCell, Cell
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import pandas as pd
import numpy as np
import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import threading
import xlsxwriter

# normalize_id consolidated below

# --- NDT Constants & Patterns ---
NAN_PATTERN = re.compile(r'^nan(\.0+)?$|^none$|^null$|^0\.0+|-0\.0+$', re.IGNORECASE)
DOT_ZERO_PATTERN = re.compile(r'\.0$')

# [PT] SCH -> 두께(mm) 변환 테이블
SCH_TO_THK = {
    "1/2": {"5S": 1.65, "10S": 2.11, "40": 2.77, "80": 3.73, "160": 4.75, "XXS": 7.47},
    "3/4": {"5S": 1.65, "10S": 2.11, "40": 2.87, "80": 3.91, "160": 5.56, "XXS": 7.82},
    "1": {"5S": 1.65, "10S": 2.77, "40": 3.38, "80": 4.55, "160": 6.35, "XXS": 9.09},
    "1-1/4": {"5S": 1.65, "10S": 2.77, "40": 3.56, "80": 4.85, "160": 6.35, "XXS": 9.70},
    "1-1/2": {"5S": 1.65, "10S": 2.77, "40": 3.68, "80": 5.08, "160": 7.14, "XXS": 10.16},
    "2": {"5S": 1.65, "10S": 2.77, "40": 3.91, "80": 5.54, "160": 8.74, "XXS": 11.07},
    "2-1/2": {"5S": 2.11, "10S": 3.05, "40": 5.16, "80": 7.01, "160": 9.53, "XXS": 14.02},
    "3": {"5S": 2.11, "10S": 3.05, "40": 5.49, "80": 7.62, "160": 11.13, "XXS": 15.24},
    "4": {"5S": 2.11, "10S": 3.05, "40": 6.02, "80": 8.56, "160": 13.49, "XXS": 17.12},
    "5": {"5S": 2.77, "10S": 3.40, "40": 6.55, "80": 9.53, "120": 12.70, "160": 15.88},
    "6": {"5S": 2.77, "10S": 3.40, "40": 7.11, "80": 10.97, "120": 14.27, "160": 18.26, "XXS": 21.95},
    "8": {"5S": 2.77, "10S": 3.76, "20": 6.35, "30": 7.04, "40": 8.18, "60": 10.31, "80": 12.70, "100": 15.09, "120": 18.26, "140": 20.62, "160": 23.01, "XXS": 22.23},
    "10": {"5S": 3.40, "10S": 4.19, "20": 6.35, "30": 7.80, "40": 9.27, "60": 12.70, "80": 15.09, "100": 18.26, "120": 21.44, "140": 25.40, "160": 28.58},
    "12": {"5S": 3.96, "10S": 4.57, "20": 6.35, "30": 8.38, "40": 10.31, "60": 14.27, "80": 17.48, "100": 21.44, "120": 25.40, "140": 28.58, "160": 33.32},
    "14": {"5S": 3.96, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 9.53, "40": 11.13, "60": 15.09, "80": 19.05, "100": 23.83, "120": 27.79, "140": 31.75, "160": 35.71},
    "16": {"5S": 4.19, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 9.53, "40": 12.70, "60": 16.66, "80": 21.44, "100": 26.19, "120": 30.96, "140": 36.53, "160": 40.49},
    "18": {"5S": 4.19, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 11.13, "40": 14.27, "60": 19.05, "80": 23.83, "100": 29.36, "120": 34.93, "140": 39.67, "160": 45.24},
    "20": {"5S": 4.78, "10S": 5.54, "10": 6.35, "20": 9.53, "30": 12.70, "40": 15.09, "60": 20.62, "80": 26.19, "100": 32.54, "120": 38.10, "140": 44.45, "160": 50.01},
    "24": {"5S": 5.54, "10S": 6.35, "10": 6.35, "20": 9.53, "30": 14.27, "40": 17.48, "60": 24.61, "80": 30.96, "100": 38.89, "120": 46.02, "140": 52.37, "160": 59.54},
}

def convert_sch_to_thk(size_val, thk_val):
    """SCH 값을 두께(mm)로 변환"""
    if pd.isna(thk_val) or str(thk_val).strip() == "": return ""
    thk_str = str(thk_val).strip().upper()
    try:
        val = float(thk_str.replace("MM", "").replace("T", "").strip())
        if 0 < val < 100: return f"{val:.2f}"
    except: pass
    sch_match = re.search(r'(?:SCH[.\s]?|S/)?(\d+S?|XXS|XS)', thk_str, re.IGNORECASE)
    if not sch_match: return thk_str
    sch = sch_match.group(1).upper()
    if sch.endswith('S') and sch not in ['5S', '10S', 'XXS', 'XS']: sch = sch[:-1]
    if pd.isna(size_val) or str(size_val).strip() == "": return thk_str
    size_str = str(size_val).strip().replace('"', '').replace("'", "")
    size_str = re.sub(r'\s+', '-', size_str)
    if size_str in SCH_TO_THK and sch in SCH_TO_THK[size_str]:
        return f"{SCH_TO_THK[size_str][sch]:.2f}"
    try:
        size_int = str(int(float(size_str)))
        if size_int in SCH_TO_THK and sch in SCH_TO_THK[size_int]:
            return f"{SCH_TO_THK[size_int][sch]:.2f}"
    except: pass
    return thk_str

# 공통 스타일 (테두리 등)
thin_side = Side(style='thin')
medium_side = Side(style='medium')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

class DailyWorkReportManager:
    def __init__(self, template_path):
        self.template_path = template_path
        # Default mapping for fallback
        self.default_mapping = {
            'header': {'date': 'F2'},
            'general': {
                'company': 'E5', 'project_name': 'E6', 'standard': 'E7', 'equipment': 'E8',
                'report_no': 'K5', 'inspection_item': 'K6', 'inspector': 'K7', 'inspector_n8': 'N8', 'car_no': 'N9'
            },
            'methods': {
                'RT': {'row': 13}, 'UT': {'row': 14}, 'MT': {'row': 15}, 'PT': {'row': 16},
                'HT': {'row': 17}, 'VT': {'row': 18}, 'LT': {'row': 19}, 'ET': {'row': 20},
                'PAUT': {'row': 21}
            },
            'rtk': {
                'center_miss': 'D34', 'density': 'F34', 'marking_miss': 'H34', 'film_mark': 'J34',
                'handling': 'L34', 'customer_complaint': 'N34', 'etc': 'P34', 'total': 'R34'
            },
            'ot': {
                'row1_name': 'B38', 'row1_company': 'F38', 'row1_method': 'I38', 'row1_hours': 'K38', 'row1_amount': 'N38',
                'row2_name': 'B39', 'row2_company': 'F39', 'row2_method': 'I39', 'row2_hours': 'K39', 'row2_amount': 'N39'
            },
            'slim_widths': {
                'A': 1, 'B': 5, 'C': 1, 'D': 6, 'E': 2, 'F': 8, 'G': 1, 'H': 5, 
                'I': 4, 'J': 4, 'K': 5, 'L': 4, 'M': 5, 'N': 4, 'O': 5, 
                'P': 4, 'Q': 5, 'R': 4, 'S': 7
            },
            'materials': {
                'RT T200': 43, 'RT AA400': 44, 'RT Other': 45,
                'MT WHITE': 46, 'MT 7C-BLACK': 47,
                'PT Penetrant': 48, 'PT Cleaner': 49, 'PT Developer': 50
            },
            'vehicles': {
                'row_start': 26, 
                'col_map': {'vehicle_info': 'B', 'mileage': 'H', 'fuel': 'C', 'clean': 'E', 'oil': 'G', 'tire': 'I', 'light': 'K', 'safety': 'M'}
            }
        }

    def generate_report(self, data, output_path, custom_mapping=None):
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import re as _re
        from copy import copy
        import openpyxl.utils
        from openpyxl.cell.cell import MergedCell, Cell

        custom_mapping = custom_mapping or {}
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template not found at {self.template_path}")

        wb = openpyxl.load_workbook(self.template_path)
        sheet = wb.active
        mapping = custom_mapping if custom_mapping else self.default_mapping
        
        black_font = Font(name='맑은 고딕', size=10, bold=False, color='000000')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
        thin = Side(style='thin')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Helper 1: free a cell by unmerging (used BEFORE merge_cells)
        def free_cell(sheet, row, col):
            for m_range in list(sheet.merged_cells.ranges):
                if row >= m_range.min_row and row <= m_range.max_row and col >= m_range.min_col and col <= m_range.max_col:
                    sheet.unmerge_cells(m_range.coord)
            key = (row, col)
            if key not in sheet._cells or isinstance(sheet._cells[key], MergedCell):
                sheet._cells[key] = Cell(sheet, row=row, column=col)

        # Helper 2: write to cell via _cells dict WITHOUT unmerging (used AFTER merge_cells)
        def write_cell(sheet, row, col, value, font=None, align=None, border=None):
            key = (row, col)
            if key not in sheet._cells or isinstance(sheet._cells[key], MergedCell):
                sheet._cells[key] = Cell(sheet, row=row, column=col)
            cell = sheet._cells[key]
            cell.value = value
            if border: cell.border = border
            elif not cell.border or cell.border == Border(): cell.border = thin_border
            if font: cell.font = font
            if align: cell.alignment = align

        def safe_write(cell_coord, value, is_currency=False, is_bold=False):
            if not cell_coord: return
            try:
                cell = sheet[cell_coord]
                if isinstance(cell, MergedCell):
                    for m_range in sheet.merged_cells.ranges:
                        if cell_coord in m_range:
                            cell = sheet.cell(row=m_range.min_row, column=m_range.min_col)
                            break
                cell.value = value
                cell.font = Font(name='맑은 고딕', size=9, bold=is_bold)
                # [FIX] shrinkToFit and wrapText are mutually exclusive in Excel - use wrapText only
                cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                if is_currency:
                    cell.number_format = '#,##0 "원"'
            except: pass

        # 1. Header & General
        d = data.get('date', datetime.date.today())
        weekdays = ["월", "화", "수", "목", "금", "토", "일"]
        date_str = f"{d.year}년 {d.month}월 {d.day}일 ({weekdays[d.weekday()]})"
        safe_write(mapping.get('header', {}).get('date', 'F2'), date_str)

        gen_map = mapping.get('general', {})
        for key in ['company', 'project_name', 'standard', 'equipment', 'report_no', 'inspection_item', 'inspector', 'car_no', 'inspector_n8']:
            val_key = 'inspector' if key == 'inspector_n8' else key
            safe_write(gen_map.get(key), data.get(val_key, ''))

        # 2. Methods (Dynamic Offset)
        methods_data = data.get('methods', {})
        current_methods = list(methods_data.keys())
        method_offset = max(0, len(current_methods) - 4)
        
        import re as _re
        def fix_c(coord, offset):
            if not coord or not isinstance(coord, str): return coord
            match = _re.match(r"([A-Z]+)([0-9]+)", coord)
            if match:
                col, row = match.groups()
                row_val = int(row)
                if row_val >= 18: return f"{col}{row_val + offset}"
            return coord

        method_base_rows = [13, 14, 15, 16]
        if method_offset > 0:
            sheet.insert_rows(17, method_offset)
            for i in range(method_offset):
                target_r = 18 + method_offset + i
                if target_r <= 25 + method_offset:
                    sheet.row_dimensions[target_r].height = 5
            method_base_rows += list(range(17, 17 + method_offset))

        # Clear default method rows (13-16) first to ensure only saved data is shown
        for r_clear in [13, 14, 15, 16]:
            for c_clear in ['B', 'E', 'H', 'K', 'N', 'Q']:
                safe_write(f"{c_clear}{r_clear}", "")

        for idx, m_name in enumerate(current_methods):
            if idx < len(method_base_rows):
                r = method_base_rows[idx]; m = methods_data[m_name]
                safe_write(f'B{r}', m_name); safe_write(f'E{r}', m.get('unit', ''))
                safe_write(f'H{r}', m.get('qty', 0)); safe_write(f'K{r}', m.get('price', 0), is_currency=True)
                safe_write(f'N{r}', m.get('travel', 0), is_currency=True); safe_write(f'Q{r}', m.get('total', 0), is_currency=True)

        # 3. Section 2 (Notes) - Styling & Normalize
        h10 = sheet.row_dimensions[10].height if sheet.row_dimensions[10].height else 15
        sheet.row_dimensions[17].height = h10
        # 4. Section 4 (RTK Quality Defect Rate)
        rtk_header_row = None
        for r_search in range(25, 45): # Target common area
            for c_search in range(1, 20):
                cell_val = str(sheet.cell(row=r_search, column=c_search).value or "")
                if any(k in cell_val for k in ["품질", "RTK"]):
                    rtk_header_row = r_search
                    print(f"DEBUG: RTK Header found at row {rtk_header_row} (Col {c_search}, Val: '{cell_val}')")
                    break
            if rtk_header_row: break
        
        if rtk_header_row:
            rtk_label_row = rtk_header_row + 1
            found_label = False
            for r_chk in range(rtk_header_row, rtk_header_row + 5):
                for c_chk in range(1, 10):
                    if "센터미스" in str(sheet.cell(row=r_chk, column=c_chk).value or ""):
                        rtk_label_row = r_chk
                        found_label = True; break
                if found_label: break
            
            target_r = rtk_label_row + 1
            rtk_data = data.get('rtk', {})
            rtk_cols = {'센터미스': 4, '농도': 6, '마킹미스': 8, '필름마크': 10, '취급부주의': 12, '고객불만': 14, '기타': 16}
            for kor, col_idx in rtk_cols.items():
                val = rtk_data.get(kor, 0)
                write_cell(sheet, target_r, col_idx, val, black_font, center_align)
            
            rtk_total = data.get('rtk_total', rtk_data.get('총계', 0))
            display_total = f"{rtk_total}매" if rtk_total else "0매"
            write_cell(sheet, target_r, 18, display_total, black_font, center_align)
            write_cell(sheet, target_r, 19, display_total, black_font, center_align)
        
        # 5. Section 5 (OT)
        ot_header_row = None
        for r_search in range(1, 100):
            if sheet.cell(row=r_search, column=2).value == "검사자":
                ot_header_row = r_search; break
        
        if ot_header_row is None:
            ot_header_row = 37 + method_offset
        else:
            if ot_header_row != 37 + method_offset:
                ot_header_row = 37 + method_offset

        ot_list = data.get('ot_status', [])
        grouped_ot = {}
        for ot in ot_list:
            comp = str(ot.get('company', '')).strip()
            meth = str(ot.get('method', '')).strip()
            raw_h = str(ot.get('ot_hours', '')).replace('₩', '').replace(',', '').strip()
            clean_h = _re.sub(r'\(.*?\)', '', raw_h).strip()
            amt = str(ot.get('ot_amount', '')).strip()
            key = (comp, meth, clean_h, amt)
            if key not in grouped_ot:
                grouped_ot[key] = {'names': [], 'company': comp, 'method': meth, 'ot_hours': raw_h, 'ot_amount': amt}
            name = str(ot.get('names', '')).strip()
            if name and name not in grouped_ot[key]['names']:
                grouped_ot[key]['names'].append(name)
        
        final_ot_list = []
        seen_ot_workers_prep = set()
        for key in grouped_ot:
            item = grouped_ot[key]
            # 이미 처리된 작업자를 제외한 새로운 인원만 필터링
            new_names = [n for n in item['names'] if n not in seen_ot_workers_prep]
            if new_names:
                item['names'] = new_names
                item['names_display'] = ", ".join(new_names)
                for n in new_names:
                    seen_ot_workers_prep.add(n)
                final_ot_list.append(item)

        ot_count = len(final_ot_list)
        ot_extra = max(0, ot_count - 2)
        if ot_extra > 0:
            print(f"DEBUG: Inserting {ot_extra} OT rows and compressing Remarks section")
            # 1. Insert rows after 39
            for i in range(ot_extra):
                new_row_idx = 40 + method_offset + i
                source_row_idx = 39 + method_offset
                sheet.insert_rows(new_row_idx)
                
                # Copy style (Font, Alignment, Border, Fill) from source row
                import copy as _copy
                for c_idx in range(2, 20): # B to S
                    source_cell = sheet.cell(row=source_row_idx, column=c_idx)
                    target_cell = sheet.cell(row=new_row_idx, column=c_idx)
                    if source_cell.has_style:
                        target_cell.font = _copy.copy(source_cell.font)
                        target_cell.alignment = _copy.copy(source_cell.alignment)
                        target_cell.border = _copy.copy(source_cell.border)
                        target_cell.fill = _copy.copy(source_cell.fill)
                
                # Copy row height too
                sheet.row_dimensions[new_row_idx].height = sheet.row_dimensions[source_row_idx].height
                
                # [DYNAMIC MERGE COPY] Re-apply all merges from the source row to the new row
                for m_range in list(sheet.merged_cells.ranges):
                    if m_range.min_row == source_row_idx and m_range.max_row == source_row_idx:
                        new_range_coord = f"{openpyxl.utils.get_column_letter(m_range.min_col)}{new_row_idx}:{openpyxl.utils.get_column_letter(m_range.max_col)}{new_row_idx}"
                        try: sheet.merge_cells(new_range_coord)
                        except: pass

            # 2. Compress Remarks section (18-25) by reducing row heights
            for i in range(min(ot_extra, 8 - method_offset)):
                target_r = 18 + method_offset + i
                sheet.row_dimensions[target_r].height = 5
        
        base_shift = method_offset + ot_extra
        for i in range(2 + ot_extra):
            r = ot_header_row + 1 + i
            for col_let in ['B', 'F', 'I', 'K', 'N', 'S']:
                safe_write(f"{col_let}{r}", '')
        
        total_ot_hours = 0.0
        for i, ot in enumerate(final_ot_list):
            r = ot_header_row + 1 + i
            
            raw_h = ot.get('ot_hours', '')
            clean_h = _re.sub(r'\(.*?\)', '', raw_h).strip()
            try: h_val = float(clean_h)
            except: h_val = clean_h
            
            ot_amount = ot.get('ot_amount', '')
            worker_name_display = ot.get('names_display', '')
            
            if isinstance(h_val, float):
                total_ot_hours += h_val
            
            safe_write(f"B{r}", worker_name_display)
            safe_write(f"F{r}", ot.get('company', ''))
            safe_write(f"I{r}", ot.get('method', ''))
            safe_write(f"K{r}", h_val)
            safe_write(f"N{r}", ot_amount, is_currency=True)
            
        # 6. Section 6 (Materials)
        materials_data = data.get('materials', {}); active_rt = []
        for k, v in materials_data.items():
            if isinstance(v, dict) and (k.startswith('RT ') or k.startswith('RT_ROW_') or v.get('is_rt')):
                name = v.get('name', '').strip()
                if not name: continue
                item = v.copy(); item['name'] = name; active_rt.append(item)
        
        rt_count = len(active_rt)
        base_rt_limit = 3
        rt_extra = max(0, rt_count - base_rt_limit)
        
        if rt_extra > 0:
            sheet.insert_rows(46, rt_extra)
            for r_new in range(46, 46 + rt_extra):
                sheet.row_dimensions[r_new].height = 30
                for col in range(1, 20):
                    try:
                        source_cell = sheet.cell(row=45, column=col)
                        target_cell = sheet.cell(row=r_new, column=col)
                        if source_cell.has_style:
                            target_cell.font = copy(source_cell.font)
                            target_cell.border = copy(source_cell.border)
                            target_cell.fill = copy(source_cell.fill)
                            target_cell.alignment = copy(source_cell.alignment)
                            target_cell.number_format = source_cell.number_format
                    except: pass
        
        total_offset = base_shift + rt_extra
        def safe_merge(sheet, s_row, s_col, e_row, e_col, value):
            for m_range in list(sheet.merged_cells.ranges):
                if not (e_row < m_range.min_row or s_row > m_range.max_row or 
                        e_col < m_range.min_col or s_col > m_range.max_col):
                    try: sheet.unmerge_cells(m_range.coord)
                    except: pass
            try:
                sheet.merge_cells(start_row=s_row, start_column=s_col, end_row=e_row, end_column=e_col)
                cell = sheet.cell(row=s_row, column=s_col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
            except: pass

        rt_display_count = max(1, rt_count)
        rt_start_row = 43 + base_shift
        rt_end_row = (rt_start_row - 1) + rt_display_count
        
        safe_merge(sheet, rt_start_row, 2, rt_end_row, 3, "RT") 
        safe_merge(sheet, 46 + total_offset, 2, 47 + total_offset, 3, "MT")
        safe_merge(sheet, 48 + total_offset, 2, 50 + total_offset, 3, "PT")

        new_sheet.column_dimensions['M'].width = 18.0
        return new_sheet

    def convert_sch_to_thk(self, size_val, thk_val):
        """SCH 값을 두께(mm)로 변환 (글로벌 함수 호출)"""
        return convert_sch_to_thk(size_val, thk_val)

    def col_to_num(self, col_str):
        """엑셀 컬럼 문자(A, B...)를 숫자(1, 2...)로 변환"""
        if not col_str: return 0
        try:
            if str(col_str).isdigit(): return int(col_str)
            return column_index_from_string(str(col_str).upper())
        except: return 0

    def is_consumable(self, name):
        """소모품(필름, 약품 등) 여부 판별"""
        if not name or pd.isna(name): return False
        n = str(name).upper().strip()
        if any(k in n for k in ["PAUT", "UT", "MPI", "PMI", "SCANNER", "WEDGE", "PROBE", "CABLE", "장비", "본체"]): return False
        if any(k in n for k in ["필름", "FILM", "CARESTREAM", "FUJIFILM", "AGFA", "KODAK", "AA400", "M100", "MX125", "T200", "HS800"]): return True
        if any(k in n for k in ["자분", "페인트", "침투제", "세척제", "현상제", "CHEMICAL", "DEVELOPER", "CLEANER", "PENETRANT", "SM-15", "MP-35", "MEGA-CHECK"]): return True
        return False

    def find_image_smart(self, keyword, exclude_keyword=None):
        def _search_in_folder(folder_path):
            if not folder_path or not os.path.exists(folder_path): return None
            candidates = glob.glob(os.path.join(folder_path, "*.*"))
            valid_extensions = ['.PNG', '.JPG', '.JPEG', '.BMP', '.GIF']
            for path in candidates:
                fname = os.path.basename(path).upper(); ext = os.path.splitext(path)[1].upper()
                if ext not in valid_extensions: continue 
                if keyword.upper() in fname:
                    if exclude_keyword and exclude_keyword.upper() in fname: continue
                    return path 
            return None
        found = _search_in_folder(self.logo_folder_path.get())
        if found: return found
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            found = _search_in_folder(sys._MEIPASS)
            if found: return found
        return None

    def place_image_freely(self, ws, img_path, anchor_cell_str, w, h, x_offset, y_offset):
        """Place image in openpyxl with precise pixel offsets and size (EMU)"""
        try:
            if not img_path or not os.path.exists(img_path): return
            from openpyxl.drawing.image import Image
            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker, XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            from openpyxl.utils import column_index_from_string
            import re
            
            img = Image(img_path)
            # 1. Calculate size in pixels
            final_w = int(w) if w > 0 else img.width
            final_h = int(h) if h > 0 else img.height
            
            # 2. Parse anchor
            m = re.match(r"([A-Z]+)([0-9]+)", anchor_cell_str)
            if not m: return
            col_str, row_str = m.groups()
            col = column_index_from_string(col_str) - 1
            row = int(row_str) - 1
            
            # 3. Create precise marker (Offset) and size (Extent) in EMU
            marker = AnchorMarker(col=col, colOff=pixels_to_EMU(x_offset), 
                                 row=row, rowOff=pixels_to_EMU(y_offset))
            # [CRITICAL] Must set extent for OneCellAnchor to show up!
            size = XDRPositiveSize2D(cx=pixels_to_EMU(final_w), cy=pixels_to_EMU(final_h))
            
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(img)
            self.log(f"   ? 로고 삽입 완료: {os.path.basename(img_path)} at {anchor_cell_str} ({final_w}x{final_h}px, Offset: {x_offset},{y_offset})")
        except Exception as e:
            self.log(f"   ? 로고 삽입 실패: {e}")

    def add_logos_to_sheet(self, ws, is_cover=False, clear_existing=True):
        # [SUPER SAFE] Identify if this is the first sheet (Cover/Gapji) of the workbook
        try: is_first_sheet = (ws == ws.parent.worksheets[0])
        except: is_first_sheet = is_cover

        # [FIX] Never clear existing images on the Cover sheet or in RT mode to preserve template diagrams
        try:
            tab_idx = self.notebook.index("current")
            is_rt = (tab_idx == 5)
        except: is_rt = False

        if is_first_sheet or is_cover or is_rt:
            clear_existing = False

        # [NEW] If RT mode and images already exist, skip adding logos to preserve the template perfectly
        if is_rt and getattr(ws, '_images', None):
            self.log(f"   ℹ️ [RT] 시트에 이미 그림(도면)이 있어 추가 로고 삽입을 건너뜁니다.")
            return

        if clear_existing:
            try: ws._images = [] 
            except: pass
        
        # [FIX] Correct tab mapping for Final version
        # Tabs: 0:Dashboard, 1:Resources, 2:Registration, 3:Materials, 4:PMI, 5:RT, 6:PT, 7:PAUT, 8:Photo Log
        mode = "RT"
        try:
            active_idx = self.notebook.index("current")
            mapping = {4: "PMI", 5: "RT", 6: "PT", 7: "PAUT", 8: "PHOTO"}
            mode = mapping.get(active_idx, "RT")
        except: pass
        
        ctx = "COVER" if is_cover else "DATA"
        context = f"{mode}_{ctx}" # [FIX] Define context here for inner functions
        self.log(f"?? {context} 로고 삽입 시작...")
        
        def _get_val(prefix, suffix, default):
            # 1. Try exact match (e.g. SITCO_RT_COVER_X)
            v = self.config.get(f"{prefix}_{context}{suffix}")
            if v is not None and v != "": return v
            # 2. Try global fallback (e.g. SITCO_COVER_X)
            fallback_ctx = "COVER" if "COVER" in context else "DATA"
            return self.config.get(f"{prefix}_{fallback_ctx}{suffix}", default)

        def _get_effective_path(prefix, keywords):
            path = self.config.get(f"{prefix}_{context}_PATH")
            if path == "": return None # Explicitly cleared
            if path and os.path.exists(path): return path
            
            if path is None:
                for kw in keywords:
                    found = self.find_image_smart(kw)
                    if found: return found
            return None

        # 1. SITCO
        p = _get_effective_path("SITCO", ["SITCO"])
        if p: self.place_image_freely(ws, p, _get_val("SITCO", "_ANCHOR", "A1"), float(_get_val("SITCO", "_W", 100)), float(_get_val("SITCO", "_H", 50)), float(_get_val("SITCO", "_X", 0)), float(_get_val("SITCO", "_Y", 0)))
        
        # 2. SEOUL
        p = _get_effective_path("SEOUL", ["서울검사"])
        if p: self.place_image_freely(ws, p, _get_val("SEOUL", "_ANCHOR", "A1"), float(_get_val("SEOUL", "_W", 100)), float(_get_val("SEOUL", "_H", 50)), float(_get_val("SEOUL", "_X", 0)), float(_get_val("SEOUL", "_Y", 0)))
        
        # 3. FOOTER
        p = _get_effective_path("FOOTER", ["바닥글", "PMI"])
        if p: self.place_image_freely(ws, p, _get_val("FOOTER", "_ANCHOR", "A1"), float(_get_val("FOOTER", "_W", 100)), float(_get_val("FOOTER", "_H", 50)), float(_get_val("FOOTER", "_X", 0)), float(_get_val("FOOTER", "_Y", 0)))

        # 4. FOOTER_PT (Left)
        p = _get_effective_path("FOOTER_PT", ["PMI갑", "PMI-1", "PT"])
        if p: self.place_image_freely(ws, p, _get_val("FOOTER_PT", "_ANCHOR", "A1"), float(_get_val("FOOTER_PT", "_W", 100)), float(_get_val("FOOTER_PT", "_H", 50)), float(_get_val("FOOTER_PT", "_X", 0)), float(_get_val("FOOTER_PT", "_Y", 0)))

    def force_print_settings(self, ws, context="DATA"):
        try:
            # [FIX] Correct tab mapping for Final version
            mode = "RT"
            try:
                active_idx = self.notebook.index("current")
                mapping = {4: "PMI", 5: "RT", 6: "PT", 7: "PAUT", 8: "PHOTO"}
                mode = mapping.get(active_idx, "RT")
            except: pass
            
            def _get_val(suffix, default):
                # Try Mode-specific first
                v = self.config.get(f"MARGIN_{mode}_{context}_{suffix}")
                if v is not None and v != "": return v
                # Fallback to general
                return self.config.get(f"MARGIN_{context}_{suffix}", default)

            ws.page_setup.paperSize = 9; ws.page_setup.orientation = 'portrait'
            
            scale = self.config.get(f"PRINT_SCALE_{mode}_{context}")
            if scale is None or scale == "": scale = self.config.get(f"PRINT_SCALE_{context}", 95)
            ws.page_setup.scale = int(float(scale))
            
            ws.print_options.horizontalCentered = True; ws.print_options.verticalCentered = True
            ws.page_margins.top = float(_get_val("TOP", 0.2))
            ws.page_margins.bottom = float(_get_val("BOTTOM", 0.2))
            ws.page_margins.left = float(_get_val("LEFT", 0.5))
            ws.page_margins.right = float(_get_val("RIGHT", 0.3))
            
            # [NEW] Manual Print Area Override
            manual_area = self.config.get(f'PRINT_AREA_{mode}_{context}', "")
            if not manual_area: manual_area = self.config.get(f'PRINT_AREA_{context}', "")
            if manual_area and manual_area.strip():
                ws.print_area = manual_area.strip()
                self.log(f"?? 인쇄 영역 설정됨: {manual_area}")
        except Exception as e:
            print(f"Print settings failed: {e}")

    def apply_custom_dimensions(self, ws, context):
        try:
            row_range_str = self.config.get(f"CUSTOM_ROWS_{context}", "").strip()
            if row_range_str:
                height = float(self.config.get(f"CUSTOM_ROW_HEIGHT_{context}", 16.5))
                for part in row_range_str.split(','):
                    if '-' in part:
                        start, end = map(int, part.split('-'))
                        for r in range(start, end + 1): ws.row_dimensions[r].height = height
                    else: ws.row_dimensions[int(part)].height = height
            col_range_str = self.config.get(f"CUSTOM_COLS_{context}", "").strip()
            if col_range_str:
                width = float(self.config.get(f"CUSTOM_COL_WIDTH_{context}", 10.0))
                for part in col_range_str.split(','):
                    if '-' in part:
                        start_l, end_l = part.split('-')
                        for c in range(column_index_from_string(start_l), column_index_from_string(end_l) + 1):
                            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = width
                    else: ws.column_dimensions[part.strip().upper()].width = width
        except: pass

    def safe_set_value(self, ws, coord, value, align=None):
        try:
            if hasattr(coord, 'coordinate'): target_cell = coord; coord_str = target_cell.coordinate
            else: target_cell = ws[coord]; coord_str = coord
            if isinstance(target_cell, MergedCell):
                for m_range in ws.merged_cells.ranges:
                    if coord_str in m_range:
                        target_cell = ws.cell(row=m_range.min_row, column=m_range.min_col); break
            target_cell.value = value
            if align: target_cell.alignment = Alignment(horizontal=align, vertical='center')
        except: pass

    def safe_merge_cells(self, ws, start_row, start_column, end_row, end_column):
        try: ws.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
        except: pass

    def set_eulji_headers(self, ws):
        headers = ["NI", "CR", "MO"]; data_font = Font(size=9); header_row = self.config.get('START_ROW', 17)
        for i, val in enumerate(headers):
            cell = ws.cell(row=header_row, column=8 + i)
            self.safe_set_value(ws, cell.coordinate, val, align='center'); cell.font = data_font
        materials = "SS304,SS304L,SS316,SS316L,SS321,SS347,SS410,SS430,DUPLEX,MONEL,INCONEL,ER308,ER308L,ER309,ER309L,ER316,ER316L,ER347,ER2209,WP316,WP316L,TP316,TP316L,F316L,A182-F316L,A312-TP316L"
        dv_q = DataValidation(type="list", formula1=f'"{materials}"', allow_blank=True); ws.add_data_validation(dv_q)
        for r in range(header_row, int(self.config.get('DATA_END_ROW', 45)) + 1):
            target_l = ws.cell(row=r, column=13); dv_q.add(target_l)
            target_l.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center'); target_l.font = Font(size=8.5)

    def prepare_next_sheet(self, wb, source_sheet_idx, page_num):
        source_sheet = wb.worksheets[source_sheet_idx]; new_sheet = wb.copy_worksheet(source_sheet) 
        base_title = source_sheet.title.split('_')[0]; new_sheet.title = f"{base_title[:20]}_{page_num:03d}"
        self.force_print_settings(new_sheet, context="DATA"); self.add_logos_to_sheet(new_sheet, is_cover=False)
        self.apply_custom_dimensions(new_sheet, "DATA")
        for col_letter, col_dim in source_sheet.column_dimensions.items(): new_sheet.column_dimensions[col_letter].width = col_dim.width
        start_row = int(self.config.get('START_ROW', 17)); end_row = int(self.config.get('DATA_END_ROW', 45))
        for r in range(start_row + 1, end_row + 1):
            for c in range(1, 14):
                cell = new_sheet.cell(row=r, column=c)
                cell.font = Font(size=8.5) if c == 13 else Font(size=9)
                self.safe_set_value(new_sheet, cell, None)
        merged_to_clear = [rng for rng in new_sheet.merged_cells.ranges if rng.min_row >= start_row and rng.max_row <= end_row]
        for rng in merged_to_clear: new_sheet.unmerge_cells(str(rng))
        return new_sheet

    def _get_val_ci(self, item, key):
        if not key or not isinstance(item, dict): return None
        if key in item: return item[key]
        k_lower = key.lower()
        for k in item.keys():
            if str(k).lower() == k_lower: return item[k]
        return None

    def check_material_grade(self, row_data):
        cr = self.to_float(self._get_val_ci(row_data, 'Cr'))
        ni = self.to_float(self._get_val_ci(row_data, 'Ni'))
        mo = self.to_float(self._get_val_ci(row_data, 'Mo'))
        mn = self.to_float(self._get_val_ci(row_data, 'Mn'))
        m = 0.1
        if (16.0*(1-m) <= cr <= 18.0*(1+m)) and (10.0*(1-m) <= ni <= 14.0*(1+m)) and (2.0*(1-m) <= mo <= 3.0*(1+m)): return "SS316"
        if (22.0*(1-m) <= cr <= 23.0*(1+m)) and (4.5*(1-m) <= ni <= 6.5*(1+m)) and (3.0*(1-m) <= mo <= 3.5*(1+m)) and (mn <= 2.2): return "DUPLEX"
        if (24.0*(1-m) <= cr <= 26.0*(1+m)) and (19.0*(1-m) <= ni <= 22.0*(1+m)): return "SS310"
        if (cr >= 16.2) and (ni >= 7.2) and (mo <= 0.55) and (mn <= 2.2): return "SS304"
        return None
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 12

        for r in range(43 + base_shift, 52 + total_offset):
            try:
                for m_range in list(sheet.merged_cells.ranges):
                    if m_range.min_row == r and m_range.max_row == r and (m_range.min_col >= 4 and m_range.max_col <= 7):
                        sheet.unmerge_cells(m_range.coord)
                sheet.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
                sheet.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
                for s_c, e_c in [(9,10), (11,12), (13,14), (15,16), (17,19)]:
                    try:
                        for m_range in list(sheet.merged_cells.ranges):
                            if m_range.min_row == r and m_range.max_row == r and m_range.min_col == s_c:
                                sheet.unmerge_cells(m_range.coord)
                        sheet.merge_cells(start_row=r, start_column=s_c, end_row=r, end_column=e_c)
                    except: pass
            except: pass

        mat_start = 43 + base_shift
        for r in range(mat_start, mat_start + 4 + rt_extra):
            for col in range(4, 20):
                cell = sheet.cell(row=r, column=col); 
                if not isinstance(cell, MergedCell): cell.value = None

        def strip_brand(name):
            for prefix in ['Carestream ', 'AGFA ', 'Fuji ', 'Kodak ']:
                if name.startswith(prefix): return name[len(prefix):]
            return name

        for idx, m in enumerate(active_rt):
            r = mat_start + idx
            if r > 46 + rt_extra: break
            display_name = strip_brand(m.get('name', ''))
            safe_write(f'D{r}', display_name)
            cell_name = sheet.cell(row=r, column=4)
            cell_name.font = Font(name='맑은 고딕', size=9)
            cell_name.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            sheet.row_dimensions[r].height = 30
            safe_write(f'F{r}', m.get('spec', ''))
            sheet.cell(row=r, column=6).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            safe_write(f'H{r}', m.get('unit', '매'))
            safe_write(f'K{r}', '-'); safe_write(f'M{r}', f"{int(m.get('used', 0))} 매")
            safe_write(f'O{r}', f"{int(m.get('in', 0))} 매")

        for r_clear in range(18 + method_offset, 26 + method_offset):
            for c_clear in range(2, 20):
                cell = sheet.cell(row=r_clear, column=c_clear)
                if not isinstance(cell, MergedCell): cell.value = None

        tiny_font = Font(size=1); no_wrap = Alignment(wrap_text=False, horizontal='center', vertical='center')
        note_range_start = 18 + method_offset; note_range_end = 25 + method_offset
        for r in range(note_range_start, note_range_end + 1):
            sheet.row_dimensions[r].height = 6.0; sheet.row_dimensions[r].custom_height = True
            for c in range(1, 20):
                cell = sheet.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.value = None; cell.alignment = no_wrap; cell.font = tiny_font
        
        merges_to_kill = [m for m in list(sheet.merged_cells.ranges) if m.min_row >= note_range_start and m.max_row <= note_range_end]
        for m in merges_to_kill:
            try: sheet.unmerge_cells(str(m))
            except: pass
            
        for r in range(13, 18 + method_offset):
            sheet.row_dimensions[r].height = 15.0; sheet.row_dimensions[r].custom_height = True
        for r in range(43 + base_shift, 43 + base_shift + rt_count):
            sheet.row_dimensions[r].height = 25.0; sheet.row_dimensions[r].custom_height = True
        title_row = 41 + base_shift; header_row = 42 + base_shift
        sheet.row_dimensions[title_row].height = 20; sheet.row_dimensions[header_row].height = 20
        sheet.row_dimensions[title_row].custom_height = True; sheet.row_dimensions[header_row].custom_height = True

        if rt_count < 4:
            for r in range(43 + base_shift + rt_count, 46 + base_shift):
                sheet.row_dimensions[r].height = 0; sheet.row_dimensions[r].hidden = True
                for col in range(1, 20):
                    cell = sheet.cell(row=r, column=col)
                    if not isinstance(cell, MergedCell):
                        cell.value = None; cell.border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

        for r in range(43 + base_shift, 43 + base_shift + rt_count):
            if r >= mat_start + len(active_rt):
                safe_write(f'K{r}', '-'); safe_write(f'M{r}', '매'); safe_write(f'O{r}', '0 매')
            
        synonyms = {
            'MT WHITE': ['MT WHITE', '백색', '백색자분', 'WHITE'],
            'MT 7C-BLACK': ['MT 7C-BLACK', '7C', '자분', 'BLACK', '7C-BLACK'],
            'PT Penetrant': ['PT Penetrant', '침투', '침투액', 'Penetrant'],
            'PT Cleaner': ['PT Cleaner', '세척', '세척액', 'Cleaner'],
            'PT Developer': ['PT Developer', '현상', '현상액', '현상제', 'Developer']
        }
        mat_map = {
            'MT WHITE': 46 + total_offset, 'MT 7C-BLACK': 47 + total_offset, 
            'PT Penetrant': 48 + total_offset, 'PT Cleaner': 49 + total_offset, 'PT Developer': 50 + total_offset
        }
        display_names = {
            'MT WHITE': '백색페인트', 'MT 7C-BLACK': '흑색자분', 
            'PT Penetrant': '침투액', 'PT Cleaner': '세척액', 'PT Developer': '현상액'
        }
        
        for m_key, r in mat_map.items():
            display_name = display_names.get(m_key, m_key)
            safe_write(f'D{r}', display_name)
            try:
                for m_range in list(sheet.merged_cells.ranges):
                    if m_range.min_row == r and m_range.max_row == r and (m_range.min_col >= 4 and m_range.max_col <= 7):
                        sheet.unmerge_cells(m_range.coord)
                sheet.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
            except: pass
            safe_write(f'H{r}', "CAN")
            m_d = {}
            for syn in synonyms.get(m_key, []):
                m_d = data.get('materials', {}).get(syn)
                if m_d: break
            if m_d:
                safe_write(f'K{r}', '-'); safe_write(f'M{r}', int(m_d.get('used', 0)))
                safe_write(f'O{r}', int(m_d.get('in', 0)))

        thin = Side(style='thin')
        for r in range(41 + base_shift, 51 + total_offset):
            for col in range(2, 20):
                cell = sheet.cell(row=r, column=col)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = center_align
        
        # [CLEANUP] Explicitly remove borders from the row after the table
        for col in range(1, 21):
            sheet.cell(row=51 + total_offset, column=col).border = Border()

        for m_key, r in mat_map.items():
            display_name = display_names.get(m_key, m_key)
            cell = sheet.cell(row=r, column=4); cell.value = display_name
            cell.font = Font(name='맑은 고딕', size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            safe_write(f'H{r}', "CAN")
            m_d = {}
            for syn in synonyms.get(m_key, []):
                m_d = data.get('materials', {}).get(syn)
                if m_d: break
            if m_d:
                safe_write(f'K{r}', '-'); safe_write(f'M{r}', int(m_d.get('used', 0)))
                safe_write(f'O{r}', int(m_d.get('in', 0)))

        sheet.row_breaks.brk = []; sheet.print_area = f'A1:S{51 + total_offset}'
        from openpyxl.worksheet.pagebreak import Break
        sheet.row_breaks.append(Break(id=51 + total_offset))
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToHeight = None; sheet.page_setup.fitToWidth = None
        sheet.sheet_properties.pageSetUpPr.fitToPage = False

        veh_list = data.get('vehicles', []); v = veh_list[0] if veh_list else {}
        veh_row = 27 + method_offset; safe_write(f"B{veh_row}", "") 
        chk_rows = {'out': 29 + total_offset, 'in': 30 + total_offset}
        for rs in range(25+total_offset, 35+total_offset):
            val = sheet.cell(row=rs, column=2).value
            if val == "출차시": chk_rows['out'] = rs
            if val == "입차시": chk_rows['in'] = rs
        
        chk_map = {'exterior':'E','cleanliness':'H','cleaning':'K','locking':'N'}
        WHITE_SQ, BLACK_SQ = "\u25a1", "\u25a0"
        import re
        for rk, row_idx in chk_rows.items():
            for ck, col_let in chk_map.items():
                val = v.get(f"{rk}_{ck}")
                if val:
                    coord = f"{col_let}{row_idx}"; cell = sheet[coord]
                    if isinstance(cell, MergedCell):
                        for mr in sheet.merged_cells.ranges:
                            if coord in mr: cell = sheet.cell(row=mr.min_row, column=mr.min_col); break
                    if cell.value and isinstance(cell.value, str):
                        pattern = f"({WHITE_SQ})(\\s*){re.escape(val)}"
                        if re.search(pattern, cell.value):
                            cell.value = re.sub(pattern, f"{BLACK_SQ}\\2{val}", cell.value)

        sheet.page_setup.horizontalCentered = True; sheet.page_setup.verticalCentered = False 
        if v.get('remarks'):
            rem_cell = f"B{31 + method_offset}"
            safe_write(rem_cell, f"비고: {v.get('remarks')}")
            if sheet[rem_cell].value:
                sheet[rem_cell].alignment = Alignment(horizontal='left', vertical='center', indent=1)
                sheet[rem_cell].font = Font(name='맑은 고딕', size=9)

        sheet.page_margins.top = 0.4; sheet.page_margins.bottom = 0.4
        sheet.page_margins.left = 0.8; sheet.page_margins.right = 0.2
        for r in range(1, 12): sheet.row_dimensions[r].height = 15
        for t_row in [12, 26, 31, 37, 42]:
            try: sheet.row_dimensions[t_row + method_offset].height = 20
            except: pass
        for v_row in [27, 28, 29, 30]: 
            try: sheet.row_dimensions[v_row + method_offset].height = 32
            except: pass
        for r in range(41 + base_shift, 52 + total_offset): sheet.row_dimensions[r].height = 18
        compact_height = 10
        for r in [5, 10, 17, 26 + method_offset, 31 + method_offset, 35 + method_offset, 40 + base_shift]:
            try: sheet.row_dimensions[r].height = compact_height
            except: pass
            
        # 8. Final Print Setup - FIXED SCALE 95%
        sheet.sheet_properties.pageSetUpPr.fitToPage = False
        sheet.page_setup.scale = 95
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.horizontalCentered = True
        sheet.page_setup.verticalCentered = False
        sheet.page_margins.top = 0.6; sheet.page_margins.bottom = 0.1
        sheet.print_area = "A1:S51"
        
        rt_start_final = 43 + base_shift; rt_end_final = (rt_start_final - 1) + max(1, rt_count)
        thin_side = Side(style='thin'); thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        title_row_idx = 41 + base_shift; header_row_idx = 42 + base_shift
        for r_fix in [title_row_idx, header_row_idx]:
            for c_idx in range(1, 20): free_cell(sheet, r_fix, c_idx)
        write_cell(sheet, title_row_idx, 1, "6.자재수불현황", Font(name='맑은 고딕', size=9, bold=True), Alignment(horizontal='left', vertical='center'))
        header_merges = [(2,3, "구분"), (4,5, "품명"), (6,7, "규격"), (8,8, "단위"), (9,10, "Lot. No."), (11,12, "입고"), (13,14, "사용"), (15,16, "잔고"), (17,19, "비고")]
        header_font = Font(name='맑은 고딕', size=9, bold=True)
        for sc, ec, label in header_merges:
            if sc != ec: sheet.merge_cells(start_row=header_row_idx, start_column=sc, end_row=header_row_idx, end_column=ec)
            write_cell(sheet, header_row_idx, sc, label, header_font, center_align)
            for c_idx in range(sc, ec + 1): sheet.cell(row=header_row_idx, column=c_idx).border = thin_border
        
        merge_plan = [
            (rt_start_final, 2, rt_end_final, 3, "RT"), 
            (46 + total_offset, 2, 47 + total_offset, 3, "MT"), 
            (48 + total_offset, 2, 50 + total_offset, 3, "PT")
        ]
        for s_row, s_col, e_row, e_col, label in merge_plan:
            for r in range(s_row, e_row + 1):
                for c in range(s_col, e_col + 1): free_cell(sheet, r, c)
            try: sheet.merge_cells(start_row=s_row, start_column=s_col, end_row=e_row, end_column=e_col)
            except: pass
            write_cell(sheet, s_row, s_col, label, black_font, center_align)
            for r in range(s_row, e_row + 1):
                for c in range(s_col, e_col + 1):
                    key = (r, c)
                    if key not in sheet._cells or isinstance(sheet._cells[key], MergedCell): sheet._cells[key] = Cell(sheet, row=r, column=c)
                    sheet._cells[key].border = thin_border
        
        for m_key, r in mat_map.items():
            write_cell(sheet, r, 4, display_names.get(m_key, m_key), black_font, center_align)

        def apply_section_style(sheet, s_row, s_col, e_row, e_col, b_style='thin', inner_style='hair'):
            main_s = Side(style=b_style); inner_side = Side(style=inner_style) if inner_style else Side(style=None)
            for r in range(s_row, e_row + 1):
                for c in range(s_col, e_col + 1):
                    cell = sheet.cell(row=r, column=c)
                    t = inner_side; b = inner_side; l = inner_side; ri = inner_side
                    if r == s_row: t = main_s
                    if r == e_row: b = main_s
                    if c == s_col: l = main_s
                    if c == e_col: ri = main_s
                    cell.border = Border(top=t, bottom=b, left=l, right=ri)

        s2_start = 18 + method_offset; s2_end = 25 + method_offset
        for m_range in list(sheet.merged_cells.ranges):
            if s2_start <= m_range.min_row and m_range.max_row <= s2_end:
                try: sheet.unmerge_cells(m_range.coord)
                except: pass
        for r_idx in range(s2_start, s2_end + 1):
            try: sheet.merge_cells(start_row=r_idx, start_column=2, end_row=r_idx, end_column=19)
            except: pass
        apply_section_style(sheet, s2_start, 2, s2_end, 19, 'thin', inner_style=None)
        apply_section_style(sheet, 12, 2, 16 + method_offset, 19, 'thin')
        apply_section_style(sheet, 6, 2, 9, 19, 'thin')
        v_title_row = 26 + method_offset; v_data_start = 28 + method_offset; v_data_end = 30 + method_offset
        apply_section_style(sheet, v_data_start, 2, v_data_end, 19, 'thin')
        for c_idx in range(2, 20): sheet.cell(row=v_title_row, column=c_idx).border = Border(top=thin_side) 
        
        # 31행 좌/우/하단 선 제거 (상단선만 유지하거나 테두리 전체 제거)
        v_remark_row = 31 + method_offset
        for c_idx in range(2, 20):
            c = sheet.cell(row=v_remark_row, column=c_idx)
            # 기존 위쪽 선은 유지하고 좌,우,아래는 없앰
            top_border = c.border.top if c.border else Side(style=None)
            c.border = Border(top=top_border, bottom=Side(style=None), left=Side(style=None), right=Side(style=None))
        apply_section_style(sheet, ot_header_row, 2, ot_header_row + 2 + ot_extra, 19, 'thin')
        apply_section_style(sheet, 43 + base_shift, 2, 50 + total_offset, 19, 'thin')
        for c in range(1, 20):
            # 자재수불현황 제목행(41+base_shift)은 윗선을 포함한 모든 선을 완전히 제거
            sheet.cell(row=41 + base_shift, column=c).border = Border()
        if rtk_header_row: apply_section_style(sheet, rtk_header_row + 1, 2, rtk_header_row + 2, 19, 'thin')
        for r_idx in [3, 4]:
            c = sheet.cell(row=r_idx, column=19); c.border = Border(left=c.border.left, top=c.border.top, bottom=c.border.bottom, right=thin_side)
        c5 = sheet.cell(row=5, column=19); c5.border = Border(left=c5.border.left, top=c5.border.top, bottom=c5.border.bottom, right=Side(style=None))
        
        # [RESTORED] Column Widths Setup
        slim_widths = {
            'A': 1, 'B': 5, 'C': 1, 'D': 6, 'E': 6, 'F': 6, 'G': 1, 'H': 5, 
            'I': 4, 'J': 4, 'K': 5, 'L': 4, 'M': 5, 'N': 5, 'O': 4, 
            'P': 4, 'Q': 4, 'R': 4, 'S': 5
        }
        for col, width in slim_widths.items():
            sheet.column_dimensions[col].width = width
            
        wb.save(output_path); return output_path


    # --- Integrated Verification Logic ---

    def check_material_grade(self, row_data):
        """Standard 10% margin alloy detection logic."""
        cr = to_float(_get_val_ci(row_data, 'Cr'))
        ni = to_float(_get_val_ci(row_data, 'Ni'))
        mo = to_float(_get_val_ci(row_data, 'Mo'))
        mn = to_float(_get_val_ci(row_data, 'Mn'))
        
        margin = 0.1 # 10% margin
        
        # 1. SUS 316 (Cr:16~18 / Ni:10~14 / Mo:2~3)
        if (16.0*(1-margin) <= cr <= 18.0*(1+margin)) and (10.0*(1-margin) <= ni <= 14.0*(1+margin)) and (2.0*(1-margin) <= mo <= 3.0*(1+margin)):
            return "SS316"
        
        # 2. Duplex (Cr:22~23 / Mo:3~3.5 / Ni:4.5~6.5 / Mn:2.0 or less)
        if (22.0*(1-margin) <= cr <= 23.0*(1+margin)) and (4.5*(1-margin) <= ni <= 6.5*(1+margin)) and (3.0*(1-margin) <= mo <= 3.5*(1+margin)) and (mn <= 2.2):
            return "DUPLEX"

        # 3. SUS 310 (Cr:24~26 / Ni:19~22)
        if (24.0*(1-margin) <= cr <= 26.0*(1+margin)) and (19.0*(1-margin) <= ni <= 22.0*(1+margin)):
            return "SS310"

        # 4. SUS 304 (Cr:18↑ / Ni:8↑ / Mo:0.5↓ / Mn:2.0↓)
        if (cr >= 16.2) and (ni >= 7.2) and (mo <= 0.55) and (mn <= 2.2):
            return "SS304"

        return None

    def evaluate_paut_flaw(self, t, h, l, depth, flaw_nature, mode="ECA"):
        """
        PAUT Acceptance Criteria Logic
        - Standard (ASME B31.1): Fatal flaws (Crack, LOF, IP) are rejected immediately.
        - ECA (Fracture Mechanics): Only Crack is fatal. LOF/IP are evaluated by size.
        """
        try:
            t_val = float(t); h_val = float(h); l_val = float(l); d_val = float(depth)
        except (ValueError, TypeError):
            return "Error (Invalid Dimension)", "Unknown"

        # 0. Automatic Location Determination
        s_top = d_val
        s_bottom = t_val - (d_val + h_val)
        s = min(s_top, s_bottom)
        s_limit = 0.4 * (h_val / 2)
        
        loc = "Surface" if s <= s_limit else "Subsurface"

        # 1. Immediate Rejection (Crack)
        nature_str = str(flaw_nature).strip().lower()
        if 'crack' in nature_str or '균열' in nature_str:
            return "Reject (Crack)", loc
            
        # Mode-based Rejection for LOF (LF) and IP
        if mode == "Standard":
            unacceptable_types = ['lof', 'lack of fusion', 'ip', 'incomplete penetration', 'lf']
            if any(x in nature_str for x in unacceptable_types):
                return f"Reject ({flaw_nature})", loc
        
        if l_val <= 0 or h_val <= 0 or t_val <= 0:
            return "Error (Zero/Negative Value)", loc

        # 1.1 Special Rules for 6mm <= t < 13mm
        if 6 <= t_val < 13:
            if l_val > 6.4: return f"Reject (L: {l_val} > 6.4mm)", loc
            if t_val < 10: h_surf_max, h_sub_max = 0.95, 0.96
            elif t_val < 12: h_surf_max, h_sub_max = 1.04, 1.04
            else: h_surf_max, h_sub_max = 1.13, 1.14
                
            limit = h_surf_max if loc == "Surface" else h_sub_max
            if h_val > limit: return f"Reject ({loc} h: {h_val} > {limit}mm)", loc
            return "Accept", loc

        # 1.2 Special Rules for 13mm <= t < 25.4mm
        if 13 <= t_val < 25.4:
            if l_val > 6.4: return f"Reject (L: {l_val} > 6.4mm)", loc
            actual_h_t = h_val / t_val
            allowed_h_t = 0.087 if loc == "Surface" else 0.143 
            if actual_h_t > allowed_h_t: return f"Reject ({loc} h/t: {actual_h_t:.3f} > {allowed_h_t:.3f})", loc
            return "Accept", loc
                
        # 2. Aspect Ratio (a/l) logic for t >= 25.4mm
        a_val = h_val if loc == "Surface" else h_val / 2
        aspect_ratio_a_l = a_val / l_val
        
        master_table = [
            (0.00, 0.031, 0.034), (0.05, 0.033, 0.038), (0.10, 0.036, 0.043),
            (0.15, 0.041, 0.054), (0.20, 0.047, 0.066), (0.25, 0.055, 0.078),
            (0.30, 0.064, 0.090), (0.35, 0.074, 0.103), (0.40, 0.083, 0.116),
            (0.45, 0.085, 0.129), (0.50, 0.087, 0.143)
        ]
        
        allowed_a_t = 0
        for ar_limit, surf_a_t, sub_a_t in master_table:
            if aspect_ratio_a_l <= ar_limit:
                allowed_a_t = surf_a_t if loc == 'Surface' else sub_a_t
                break
        if allowed_a_t == 0: allowed_a_t = master_table[-1][1] if loc == 'Surface' else master_table[-1][2]

        actual_a_t = a_val / t_val
        if actual_a_t <= allowed_a_t: return "Accept", loc
        return f"Reject ({loc} a/t: {actual_a_t:.3f} > {allowed_a_t:.3f})", loc

# [OPTIMIZATION] Prevent slow Windows registry scan for mimetypes
if os.name == 'nt':
    mimetypes.init(files=[])

# DPI Awareness for Windows
try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except:
    pass

warnings.simplefilter("ignore")

# --- Constants & Paths ---
NAN_PATTERN = re.compile(r'^nan(\.0+)?$|^none$|^null$|^0\.0+|-0\.0+$', re.IGNORECASE)
DOT_ZERO_PATTERN = re.compile(r'\.0$')
MARKER_PATTERN = re.compile(r'\(.*?\)\s*|익일')

# [PT] SCH -> 두께(mm) 변환 테이블
SCH_TO_THK = {
    "1/2": {"5S": 1.65, "10S": 2.11, "40": 2.77, "80": 3.73, "160": 4.75, "XXS": 7.47},
    "3/4": {"5S": 1.65, "10S": 2.11, "40": 2.87, "80": 3.91, "160": 5.56, "XXS": 7.82},
    "1": {"5S": 1.65, "10S": 2.77, "40": 3.38, "80": 4.55, "160": 6.35, "XXS": 9.09},
    "1-1/4": {"5S": 1.65, "10S": 2.77, "40": 3.56, "80": 4.85, "160": 6.35, "XXS": 9.70},
    "1-1/2": {"5S": 1.65, "10S": 2.77, "40": 3.68, "80": 5.08, "160": 7.14, "XXS": 10.16},
    "2": {"5S": 1.65, "10S": 2.77, "40": 3.91, "80": 5.54, "160": 8.74, "XXS": 11.07},
    "2-1/2": {"5S": 2.11, "10S": 3.05, "40": 5.16, "80": 7.01, "160": 9.53, "XXS": 14.02},
    "3": {"5S": 2.11, "10S": 3.05, "40": 5.49, "80": 7.62, "160": 11.13, "XXS": 15.24},
    "4": {"5S": 2.11, "10S": 3.05, "40": 6.02, "80": 8.56, "160": 13.49, "XXS": 17.12},
    "5": {"5S": 2.77, "10S": 3.40, "40": 6.55, "80": 9.53, "120": 12.70, "160": 15.88},
    "6": {"5S": 2.77, "10S": 3.40, "40": 7.11, "80": 10.97, "120": 14.27, "160": 18.26, "XXS": 21.95},
    "8": {"5S": 2.77, "10S": 3.76, "20": 6.35, "30": 7.04, "40": 8.18, "60": 10.31, "80": 12.70, "100": 15.09, "120": 18.26, "140": 20.62, "160": 23.01, "XXS": 22.23},
    "10": {"5S": 3.40, "10S": 4.19, "20": 6.35, "30": 7.80, "40": 9.27, "60": 12.70, "80": 15.09, "100": 18.26, "120": 21.44, "140": 25.40, "160": 28.58},
    "12": {"5S": 3.96, "10S": 4.57, "20": 6.35, "30": 8.38, "40": 10.31, "60": 14.27, "80": 17.48, "100": 21.44, "120": 25.40, "140": 28.58, "160": 33.32},
    "14": {"5S": 3.96, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 9.53, "40": 11.13, "60": 15.09, "80": 19.05, "100": 23.83, "120": 27.79, "140": 31.75, "160": 35.71},
    "16": {"5S": 4.19, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 9.53, "40": 12.70, "60": 16.66, "80": 21.44, "100": 26.19, "120": 30.96, "140": 36.53, "160": 40.49},
    "18": {"5S": 4.19, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 11.13, "40": 14.27, "60": 19.05, "80": 23.83, "100": 29.36, "120": 34.93, "140": 39.67, "160": 45.24},
    "20": {"5S": 4.78, "10S": 5.54, "10": 6.35, "20": 9.53, "30": 12.70, "40": 15.09, "60": 20.62, "80": 26.19, "100": 32.54, "120": 38.10, "140": 44.45, "160": 50.01},
    "24": {"5S": 5.54, "10S": 6.35, "10": 6.35, "20": 9.53, "30": 14.27, "40": 17.48, "60": 24.61, "80": 30.96, "100": 38.89, "120": 46.02, "140": 52.37, "160": 59.54},
}

def install_and_import(package, import_name=None):
    if import_name is None: import_name = package
    try:
        return __import__(import_name)
    except ImportError:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package, "--break-system-packages"])
            return __import__(import_name)
        except:
            return None

install_and_import('tkcalendar')
from tkcalendar import DateEntry, Calendar

# --- tkcalendar Patches ---
try:
    import tkinter as _tk_ref
    _orig_cal_init = Calendar.__init__
    def _patched_cal_init(self, *args, **kwargs):
        if 'select_on_nav' not in kwargs: kwargs['select_on_nav'] = False
        _orig_cal_init(self, *args, **kwargs)
        def widen():
            if hasattr(self, '_header_month'): self._header_month.configure(width=25)
            if hasattr(self, '_header_year'): self._header_year.configure(width=10)
        self.after_idle(widen)
    Calendar.__init__ = _patched_cal_init
    
    if hasattr(DateEntry, 'drop_down'):
        _orig_drop_down = DateEntry.drop_down
        def _patched_drop_down(self):
            try: _orig_drop_down(self)
            except: pass
        DateEntry.drop_down = _patched_drop_down
except: pass

# --- GLOBAL UTILITY FUNCTIONS ---
def normalize_id(val):
    """Robust ID normalization for Excel/CSV data."""
    if pd.isna(val) or val == '' or str(val).lower() == 'nan': return ""
    try:
        # Handle cases like 123.0 or "123.0"
        return str(int(float(val)))
    except:
        return str(val).strip()

def to_float(val):
    try:
        if val is None: return 0.0
        s = str(val).replace('%', '').replace(',', '').strip()
        if not s or s.lower() == 'nan': return 0.0
        return float(s)
    except: return 0.0

def _get_val_ci(item, key):
    """Case-insensitive dictionary lookup."""
    if not key or not isinstance(item, dict): return None
    if key in item: return item[key]
    k_lower = key.lower()
    for k in item.keys():
        if str(k).lower() == k_lower:
            return item[k]
    return None


# Common styles
thin_side = Side(style='thin')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Paths
BASE_DIR = os.getcwd()
DATA_DIR = os.path.join(BASE_DIR, "data")
CONFIG_DIR = os.path.join(BASE_DIR, "config")
RESOURCE_DIR = os.path.join(BASE_DIR, "resources")
for d in [DATA_DIR, CONFIG_DIR, RESOURCE_DIR]: os.makedirs(d, exist_ok=True)

SETTINGS_FILE = os.path.join(CONFIG_DIR, "logo_settings_unified.json")

class DraggableMessagebox:
    @staticmethod
    def _show(type, title, message):
        root = tk._default_root
        if not root: return
        dialog = tk.Toplevel(root)
        dialog.overrideredirect(True)
        dialog.attributes("-topmost", True)
        dialog.config(background="#f3f4f6", highlightthickness=1, highlightbackground="#d1d5db")
        title_bar = tk.Frame(dialog, background="#ffffff", height=30, cursor="fleur")
        title_bar.pack(side="top", fill="x")
        tk.Label(title_bar, text=title, font=("Malgun Gothic", 9, "bold"), background="#ffffff", padx=10).pack(side="left")
        def close_dialog(): dialog.grab_release(); dialog.destroy()
        btn_close = tk.Label(title_bar, text="✕", font=("Malgun Gothic", 10), background="#ffffff", padx=10, cursor="hand2")
        btn_close.pack(side="right")
        btn_close.bind("<Button-1>", lambda e: close_dialog())
        dialog.bind("<Button-1>", lambda e: setattr(dialog, '_drag_start', (e.x, e.y)))
        dialog.bind("<B1-Motion>", lambda e: dialog.geometry(f"+{dialog.winfo_x()+e.x-dialog._drag_start[0]}+{dialog.winfo_y()+e.y-dialog._drag_start[1]}"))
        main_frame = tk.Frame(dialog, background="#f3f4f6", padx=20, pady=20)
        main_frame.pack(expand=True, fill='both')
        tk.Label(main_frame, text=message, font=("Malgun Gothic", 10), justify="left", wraplength=480, background="#f3f4f6").pack(side="left")
        tk.Button(dialog, text="확인", command=close_dialog, width=10).pack(pady=10)
        dialog.grab_set(); root.wait_window(dialog)
    @staticmethod
    def showerror(t, m): DraggableMessagebox._show("error", t, m)
    @staticmethod
    def showwarning(t, m): DraggableMessagebox._show("warning", t, m)
    @staticmethod
    def showinfo(t, m): DraggableMessagebox._show("info", t, m)

messagebox.showerror = DraggableMessagebox.showerror
messagebox.showwarning = DraggableMessagebox.showwarning
messagebox.showinfo = DraggableMessagebox.showinfo

class WorkerCompositeWidget(ttk.Frame):
    def __init__(self, parent, enable_autocomplete=False, user_list=None, **kwargs):
        super().__init__(parent)
        name_width = kwargs.pop('width', 15)
        self.cb_name = ttk.Combobox(self, width=name_width, **kwargs)
        self.cb_name.pack(side='left', fill='x', expand=True)
    def get(self): return self.cb_name.get().strip()
    def set(self, value):
        if not value: self.cb_name.set(""); return
        match = re.match(r"\((주간|야간|휴일|주야간)\)\s*(.*)", str(value))
        if match: self.cb_name.set(match.group(2).strip())
        else: self.cb_name.set(str(value).strip())
    def bind(self, sequence=None, func=None, add=None): self.cb_name.bind(sequence, func, add)

class WorkerDataGroup(ttk.Frame):
    def __init__(self, parent, worker_index, users_list, time_list=None, enable_autocomplete=False, **kwargs):
        super().__init__(parent, padding=2)
        self.worker_index = worker_index
        self.composite = WorkerCompositeWidget(self, width=12, values=users_list)
        self.composite.pack(side='left', padx=(0, 2))
        self.cb_name = self.composite.cb_name
        self.cb_shift = ttk.Combobox(self, values=["주간", "야간", "휴일", "주야간"], width=6, state="readonly")
        self.cb_shift.pack(side='left', padx=(1, 2))
        self.cb_shift.set("주간")
        ttk.Label(self, text="시간:").pack(side='left', padx=(1, 0))
        self.ent_worktime = ttk.Combobox(self, width=16, values=time_list or [])
        self.ent_worktime.pack(side='left', padx=(0, 2))
        ttk.Label(self, text="OT:").pack(side='left', padx=(1, 0))
        self.ent_ot = ttk.Entry(self, width=18)
        self.ent_ot.pack(side='left')
    def get_worker(self): return self.composite.get()
    def set_worker(self, val): self.composite.set(val)
    def get_time(self):
        shift = self.cb_shift.get()
        time = self.ent_worktime.get().strip()
        return f"({shift}) {time}" if time else ""
    def set_time(self, val):
        if not val: self.ent_worktime.set(""); self.cb_shift.set("주간"); return
        match = re.match(r"\((주간|야간|휴일|주야간)\)\s*(.*)", str(val))
        if match: self.cb_shift.set(match.group(1)); self.ent_worktime.set(match.group(2).strip())
        else: self.cb_shift.set("주간"); self.ent_worktime.set(str(val).strip())
    def get_ot(self): return self.ent_ot.get()
    def set_ot(self, val): self.ent_ot.delete(0, tk.END); self.ent_ot.insert(0, val)

class VehicleInspectionWidget(ttk.Frame):
    def __init__(self, parent, vehicle_list=None, **kwargs):
        super().__init__(parent)
        self.vars = {}
        rows = [("출차시", "out"), ("입차시", "in")]
        categories = [("exterior", ["양호", "불량"]), ("cleanliness", ["양호", "불량"]), ("cleaning", ["함", "안함"]), ("locking", ["잠금", "안함"])]
        for r_idx, (r_label, r_key) in enumerate(rows):
            ttk.Label(self, text=r_label).grid(row=r_idx, column=0, padx=5, pady=2)
            for c_idx, (c_key, opts) in enumerate(categories):
                var = tk.StringVar()
                self.vars[f"{r_key}_{c_key}"] = var
                cb = ttk.Combobox(self, textvariable=var, values=opts, width=6)
                cb.grid(row=r_idx, column=c_idx+1, padx=2, pady=2)
        
        ttk.Label(self, text="차량번호:").grid(row=2, column=0)
        self.cb_vno = ttk.Combobox(self, values=vehicle_list or [])
        self.cb_vno.grid(row=2, column=1, columnspan=2)
        ttk.Label(self, text="주행거리:").grid(row=2, column=3)
        self.ent_mileage = ttk.Entry(self, width=10)
        self.ent_mileage.grid(row=2, column=4)

    def reset_fields(self):
        for v in self.vars.values(): v.set("")
        self.cb_vno.set(""); self.ent_mileage.delete(0, tk.END)

def enable_column_resize(frame, num_cols, header_row=0, edge_px=6):
    """Adds drag-to-resize functionality to grid columns."""
    _drag = {'col': None, 'start_x': 0, 'start_w': 0, 'next_w': 0}
    def get_col_width(col):
        try:
            info = frame.grid_columnconfigure(col)
            minsz = info.get('minsize', 0)
            if minsz and minsz > 0: return minsz
            for w in frame.grid_slaves(row=header_row, column=col):
                if w.winfo_width() > 1: return w.winfo_width()
            return 80
        except: return 80
    def on_motion(event):
        if _drag['col'] is not None: return
        w = event.widget; x = event.x; wd = w.winfo_width()
        if wd - edge_px <= x <= wd: w.configure(cursor='sb_h_double_arrow')
        else: w.configure(cursor='')
    def on_press(event):
        w = event.widget; x = event.x; wd = w.winfo_width()
        if wd - edge_px <= x <= wd:
            col = w.grid_info().get('column', -1)
            if col < 0 or col >= num_cols - 1: return
            _drag.update({'col': col, 'start_x': event.x_root, 'start_w': get_col_width(col), 'next_w': get_col_width(col + 1)})
        else: _drag['col'] = None
    def on_drag(event):
        if _drag['col'] is None: return
        dx = event.x_root - _drag['start_x']
        frame.grid_columnconfigure(_drag['col'], minsize=max(30, _drag['start_w'] + dx), weight=0)
        frame.grid_columnconfigure(_drag['col'] + 1, minsize=max(30, _drag['next_w'] - dx), weight=0)
    def bind_headers():
        for col in range(num_cols):
            for w in frame.grid_slaves(row=header_row, column=col):
                w.bind('<Motion>', on_motion, add='+'); w.bind('<Leave>', lambda e: w.configure(cursor=''), add='+')
                w.bind('<ButtonPress-1>', on_press, add='+'); w.bind('<B1-Motion>', on_drag, add='+')
                w.bind('<ButtonRelease-1>', lambda e: _drag.update({'col': None}), add='+')
    frame.after_idle(bind_headers)

class LaborCostDetailWidget(ttk.Frame):
    def __init__(self, parent, on_change_callback=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.on_change_callback = on_change_callback
        self.ranks = ["이사", "부장", "차장", "과장", "대리", "계장", "주임", "기사"]
        self.special_types = ["연장근무", "야간근무", "휴일근무"]
        self.base_salaries = {"이사": 55250000, "부장": 55250000, "차장": 47670000, "과장": 41170000, "대리": 37920000, "계장": 34670000, "주임": 31420000, "기사": 29250000}
        self.entries = {}; self.totals = {}
        self._create_widgets()
    def _create_widgets(self):
        style = ttk.Style()
        style.configure("LaborHeader.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#e0e0e0', relief='solid')
        style.configure("LaborTotal.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#fff9c4', relief='solid')
        c_f = ttk.Frame(self); c_f.pack(side='left', fill='both', expand=True)
        r_f = ttk.Frame(self, padding=(20, 40, 0, 0)); r_f.pack(side='right', fill='y')
        ttk.Label(c_f, text="1) 정시근무 (240일/년)", font=('Malgun Gothic', 11, 'bold')).pack(anchor='w', pady=(10, 5))
        t1 = ttk.Frame(c_f); t1.pack(fill='x')
        h1 = ["구분", "직급", "인원(명)", "일수(일)", "단가/일", "사전원가가액"]
        for j, h in enumerate(h1): ttk.Label(t1, text=h, style="LaborHeader.TLabel", padding=5, anchor='center').grid(row=0, column=j, sticky='nsew'); t1.grid_columnconfigure(j, weight=1)
        enable_column_resize(t1, len(h1))
        ttk.Label(t1, text="정시근무", relief='solid', anchor='center').grid(row=1, column=0, rowspan=len(self.ranks), sticky='nsew')
        for i, rank in enumerate(self.ranks):
            row = i + 1; ttk.Label(t1, text=rank, relief='solid', anchor='center').grid(row=row, column=1, sticky='nsew'); self.entries[rank] = {}
            p = ttk.Entry(t1, width=8, justify='center'); p.grid(row=row, column=2, sticky='nsew'); p.bind("<KeyRelease>", lambda e, r=rank: self._on_input_change(r)); self.entries[rank]['personnel'] = p
            d = ttk.Entry(t1, width=8, justify='center'); d.grid(row=row, column=3, sticky='nsew'); d.bind("<KeyRelease>", lambda e, r=rank: self._on_input_change(r)); self.entries[rank]['period'] = d
            u = ttk.Entry(t1, width=12, justify='right'); u.grid(row=row, column=4, sticky='nsew'); u.bind("<KeyRelease>", lambda e, r=rank: self._on_input_change(r)); self.entries[rank]['unit_price'] = u
            u.insert(0, f"{round(self.base_salaries[rank]/240):,.0f}")
            l = ttk.Label(t1, text="0", relief='solid', anchor='e', padding=5); l.grid(row=row, column=5, sticky='nsew'); self.totals[rank] = l
        ttk.Label(c_f, text="2) 특별근무", font=('Malgun Gothic', 11, 'bold')).pack(anchor='w', pady=(20, 5))
        t2 = ttk.Frame(c_f); t2.pack(fill='x')
        h2 = ["구분", "형태", "인원(명)", "시간(H)", "단가", "사전원가가액"]
        for j, h in enumerate(h2): ttk.Label(t2, text=h, style="LaborHeader.TLabel", padding=5, anchor='center').grid(row=0, column=j, sticky='nsew'); t2.grid_columnconfigure(j, weight=1)
        enable_column_resize(t2, len(h2))
        ttk.Label(t2, text="특별근무", relief='solid', anchor='center').grid(row=1, column=0, rowspan=len(self.special_types), sticky='nsew')
        for i, st in enumerate(self.special_types):
            row = i + 1; ttk.Label(t2, text=st, relief='solid', anchor='center').grid(row=row, column=1, sticky='nsew'); self.entries[st] = {}
            p = ttk.Entry(t2, width=8, justify='center'); p.grid(row=row, column=2, sticky='nsew'); p.bind("<KeyRelease>", lambda e, s=st: self._on_input_change(s)); self.entries[st]['personnel'] = p
            h = ttk.Entry(t2, width=8, justify='center'); h.grid(row=row, column=3, sticky='nsew'); h.bind("<KeyRelease>", lambda e, s=st: self._on_input_change(s)); self.entries[st]['period'] = h
            u = ttk.Entry(t2, width=12, justify='right'); u.grid(row=row, column=4, sticky='nsew'); u.bind("<KeyRelease>", lambda e, s=st: self._on_input_change(s)); self.entries[st]['unit_price'] = u
            u.insert(0, f"{({'연장근무':4000, '야간근무':5000, '휴일근무':7500}.get(st,0)):,.0f}")
            l = ttk.Label(t2, text="0", relief='solid', anchor='e', padding=5); l.grid(row=row, column=5, sticky='nsew'); self.totals[st] = l
        f_tot = ttk.Frame(c_f); f_tot.pack(fill='x', pady=10)
        ttk.Label(f_tot, text="인건비 합계", style="LaborTotal.TLabel", padding=10).pack(side='left', fill='x', expand=True)
        self.lbl_grand_total = ttk.Label(f_tot, text="₩ 0", style="LaborTotal.TLabel", anchor='e', font=('Malgun Gothic', 12, 'bold'), padding=10); self.lbl_grand_total.pack(side='right', fill='x', expand=True)
    def _on_input_change(self, key):
        p = to_float(self.entries[key]['personnel'].get()); d = to_float(self.entries[key]['period'].get()); u = to_float(self.entries[key]['unit_price'].get())
        self.totals[key].config(text=f"{p*d*u:,.0f}"); self.calculate_all()
    def calculate_all(self):
        t1 = sum(to_float(self.totals[r].cget('text')) for r in self.ranks)
        t2 = sum(to_float(self.totals[s].cget('text')) for s in self.special_types)
        gt = t1 + t2; self.lbl_grand_total.config(text=f"₩ {gt:,.0f}")
        if self.on_change_callback: self.on_change_callback(gt)
    def get_data(self): return {k: {wk: wv.get() for wk, wv in row.items()} for k, row in self.entries.items()}
    def set_data(self, data):
        if not data: return
        for k, v in data.items():
            if k in self.entries:
                for wk, val in v.items(): self.entries[k][wk].delete(0, tk.END); self.entries[k][wk].insert(0, val)
        for k in list(self.ranks) + list(self.special_types): self._on_input_change(k)

class MaterialCostDetailWidget(ttk.Frame):
    def __init__(self, parent, on_change_callback=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.on_change_callback = on_change_callback
        self.items = [("PT 약품 세척제", "CAN", 1500), ("PT 약품 침투제", "CAN", 2300), ("PT 약품 현상제", "CAN", 2000), ("RT 필름 MX125", "매", 990), ("글리세린 20L", "통", 100000)]
        self.entries = []; self._create_widgets()
    def _create_widgets(self):
        ttk.Label(self, text="2) 재료비", font=('Malgun Gothic', 11, 'bold')).pack(anchor='w', pady=(10, 5))
        t = ttk.Frame(self); t.pack(fill='x')
        headers = ["품목", "수량", "규격", "단가", "금액"]
        for j, h in enumerate(headers): ttk.Label(t, text=h, background='#e0e0e0', relief='solid', padding=5, anchor='center').grid(row=0, column=j, sticky='nsew'); t.grid_columnconfigure(j, weight=1)
        for i, (name, unit, price) in enumerate(self.items):
            row = i + 1; ttk.Label(t, text=name, relief='solid', padding=5).grid(row=row, column=0, sticky='nsew')
            q = ttk.Entry(t, width=8, justify='center'); q.grid(row=row, column=1, sticky='nsew'); q.bind("<KeyRelease>", lambda e, idx=i: self._on_change(idx))
            ttk.Label(t, text=unit, relief='solid', anchor='center').grid(row=row, column=2, sticky='nsew')
            p = ttk.Entry(t, width=10, justify='right'); p.grid(row=row, column=3, sticky='nsew'); p.insert(0, f"{price:,.0f}"); p.bind("<KeyRelease>", lambda e, idx=i: self._on_change(idx))
            a = ttk.Label(t, text="0", relief='solid', anchor='e', padding=5); a.grid(row=row, column=4, sticky='nsew')
            self.entries.append({'qty': q, 'price': p, 'amt': a})
        self.lbl_tot = ttk.Label(self, text="재료비 합계: ₩ 0", font=('Malgun Gothic', 11, 'bold'), background='#ffff00', relief='solid', padding=10, anchor='e')
        self.lbl_tot.pack(fill='x', pady=10)
    def _on_change(self, idx):
        row = self.entries[idx]; q = to_float(row['qty'].get()); p = to_float(row['price'].get())
        row['amt'].config(text=f"{q*p:,.0f}"); self.calculate_all()
    def calculate_all(self):
        gt = sum(to_float(r['amt'].cget('text')) for r in self.entries)
        self.lbl_tot.config(text=f"재료비 합계: ₩ {gt:,.0f}")
        if self.on_change_callback: self.on_change_callback(gt)
    def get_data(self): return [{'qty': r['qty'].get(), 'price': r['price'].get()} for r in self.entries]
    def set_data(self, data):
        if not data: return
        for i, v in enumerate(data):
            if i < len(self.entries):
                self.entries[i]['qty'].delete(0, tk.END); self.entries[i]['qty'].insert(0, v.get('qty', ''))
                self.entries[i]['price'].delete(0, tk.END); self.entries[i]['price'].insert(0, v.get('price', ''))
                self._on_change(i)

class ExpenseProfitDetailWidget(ttk.Frame):
    def __init__(self, parent, on_change_callback=None, get_labor_func=None, get_mat_func=None, get_rev_func=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.on_change_callback = on_change_callback
        self.get_labor = get_labor_func; self.get_mat = get_mat_func; self.get_rev = get_rev_func
        self.entries = {'exp': [], 'out': []}; self._create_widgets()
    def _create_widgets(self):
        ttk.Label(self, text="3) 경비 및 영업이익", font=('Malgun Gothic', 11, 'bold')).pack(anchor='w', pady=(10, 5))
        f1 = ttk.LabelFrame(self, text="현장경비"); f1.pack(fill='x', pady=5)
        t1 = ttk.Frame(f1); t1.pack(fill='x')
        for j, h in enumerate(["구분", "수량", "단가", "금액"]): ttk.Label(t1, text=h, background='#e0e0e0', relief='solid', padding=5, anchor='center').grid(row=0, column=j, sticky='nsew'); t1.grid_columnconfigure(j, weight=1)
        for i, (n, p) in enumerate([("차량유지비", 5000), ("소모품비", 500), ("복리후생비", 1667)]):
            r = i + 1; ttk.Label(t1, text=n, relief='solid').grid(row=r, column=0, sticky='nsew')
            q = ttk.Entry(t1, width=8, justify='center'); q.grid(row=r, column=1, sticky='nsew'); q.bind("<KeyRelease>", lambda e: self.calculate_all())
            pr = ttk.Entry(t1, width=10, justify='right'); pr.grid(row=r, column=2, sticky='nsew'); pr.insert(0, f"{p:,.0f}"); pr.bind("<KeyRelease>", lambda e: self.calculate_all())
            a = ttk.Label(t1, text="0", relief='solid', anchor='e', padding=5); a.grid(row=r, column=3, sticky='nsew')
            self.entries['exp'].append({'qty': q, 'price': pr, 'amt': a})
        self.lbl_res = ttk.Label(self, text="영업이익: ₩ 0 (0.00%)", font=('Malgun Gothic', 11, 'bold'), background='#90ee90', relief='solid', padding=10, anchor='center')
        self.lbl_res.pack(fill='x', pady=10)
    def calculate_all(self):
        te = 0
        for r in self.entries['exp']:
            a = to_float(r['qty'].get()) * to_float(r['price'].get())
            r['amt'].config(text=f"{a:,.0f}"); te += a
        l = self.get_labor() if self.get_labor else 0
        m = self.get_mat() if self.get_mat else 0
        rv = self.get_rev() if self.get_rev else 0
        tc = (l + m + te) * 1.14 # Direct + Indirect
        pr = rv - tc; pct = (pr / rv * 100) if rv > 0 else 0
        self.lbl_res.config(text=f"영업이익: ₩ {pr:,.0f} ({pct:.2f}%)")
        if self.on_change_callback: self.on_change_callback(te, 0, pr)
    def get_data(self): return {k: [{'qty': r['qty'].get(), 'price': r['price'].get()} for r in v] for k, v in self.entries.items()}
    def set_data(self, data):
        if not data: return
        for k in ['exp']:
            for i, v in enumerate(data.get(k, [])):
                if i < len(self.entries[k]):
                    self.entries[k][i]['qty'].delete(0, tk.END); self.entries[k][i]['qty'].insert(0, v.get('qty', ''))
                    self.entries[k][i]['price'].delete(0, tk.END); self.entries[k][i]['price'].insert(0, v.get('price', ''))
        self.calculate_all()

# --- Integrated Smart Manager Class ---
class IntegratedSmartManager:
    def __init__(self, root):
        self.root = root
        self.root.title("Integrated Smart Manager (Material + NDT) - v1.5.8")
        self.is_ready = False
        self.daily_usage_sash_locked = False
        self.draggable_items = {}
        
        # --- Environment & Configuration ---
        self.base_dir = os.getcwd()
        self.data_dir = os.path.join(self.base_dir, "data")
        self.config_dir = os.path.join(self.base_dir, "config")
        self.resource_dir = os.path.join(self.base_dir, "resources")
        for d in [self.data_dir, self.config_dir, self.resource_dir]:
            os.makedirs(d, exist_ok=True)

        self.db_path = os.path.join(self.data_dir, "Material_Inventory.xlsx")
        self.settings_file = os.path.join(self.config_dir, "logo_settings_unified.json")
        self.tab_config_file = os.path.join(self.config_dir, "tab_config.json")
        self.config_path = os.path.join(self.config_dir, "Material_Manager_Config.json")
        
        # --- NDT Reporting Variables ---
        self.config = self.load_settings()
        self.extracted_data = []      # PMI
        self.rt_extracted_data = []   # RT
        self.pt_extracted_data = []   # PT
        self.paut_extracted_data = [] # PAUT
        
        # Search & Filter variables
        self.pmi_show_deficiency_only = tk.BooleanVar(value=False)
        self.rt_search_var = tk.StringVar()
        self.pt_search_var = tk.StringVar()
        self.paut_search_var = tk.StringVar()
        
        # File paths (NDT)
        self.target_file_path = tk.StringVar()
        self.template_file_path = tk.StringVar()
        self.rt_target_file_path = tk.StringVar()
        self.rt_template_file_path = tk.StringVar()
        self.pt_target_file_path = tk.StringVar()
        self.pt_template_file_path = tk.StringVar()
        self.paut_target_file_path = tk.StringVar()
        self.paut_template_file_path = tk.StringVar()

        # [NEW] Add traces for Template-Linked Config Auto-Load
        self.rt_template_file_path.trace_add("write", lambda *args: self.load_template_specific_config(self.rt_template_file_path.get(), "RT"))
        self.pt_template_file_path.trace_add("write", lambda *args: self.load_template_specific_config(self.pt_template_file_path.get(), "PT"))
        self.paut_template_file_path.trace_add("write", lambda *args: self.load_template_specific_config(self.paut_template_file_path.get(), "PAUT"))
        self.template_file_path.trace_add("write", lambda *args: self.load_template_specific_config(self.template_file_path.get(), "PMI"))
        
        # File info display
        self.file_info_vars = {
            'PMI': tk.StringVar(value="📂 파일을 선택해주세요."),
            'RT': tk.StringVar(value="📂 파일을 선택해주세요."),
            'PT': tk.StringVar(value="📂 파일을 선택해주세요."),
            'PAUT': tk.StringVar(value="📂 파일을 선택해주세요.")
        }
        
        # --- Material Management Variables ---
        self.materials_df = None
        self.transactions_df = None
        self.monthly_usage_df = None
        self.daily_usage_df = None
        self.budget_df = None
        self.settings_df = None
        
        # List properties (Synchronized across UI)
        self.sites = []
        self.budget_sites = []
        self.users = []
        self.vehicles = []
        self.warehouses = []
        self.test_methods = ["RT", "UT", "MT", "PT", "PAUT", "PMI", "ECT", "LT", "VT", "ET", "FT", "HT", "ST", "Other"]
        self.co_code_list = ["SITCO", "SEOUL"]
        self.class_list = []
        self.mfr_list = []
        self.sn_list = []
        self.model_list = []
        self.eq_code_list = []
        self.item_name_list = []
        self.spec_list = []
        self.unit_list = ["EA", "CAN", "매", "통", "kg", "m", "pk", "box", "roll", "set", "Other"]
        self.supplier_list = []
        self.origin_list = []
        self.worktimes = [
            "08:00~17:00", "08:30~17:30", "09:00~18:00", "18:00~03:00", "20:00~05:00",
            "17:00~02:00", "08:00~12:00", "13:00~17:00", "08:00~18:00", "08:00~20:00",
            "08:00~22:00", "12시간 교대", "기타"
        ]
        
        # [NEW] PAUT Manual Evaluation Variables
        self.paut_manual_vars = {
            't': tk.StringVar(), 'h': tk.StringVar(), 'l': tk.StringVar(), 'd': tk.StringVar(),
            'z1': tk.StringVar(), 'z2': tk.StringVar(), 'L1': tk.StringVar(), 'L2': tk.StringVar(),
            'D_Upper': tk.StringVar(), 'D_Lower': tk.StringVar(), 'nature': tk.StringVar(value="Slag"),
            'loc': tk.StringVar(value="-"), 'peak': tk.StringVar(value="80"), 'db': tk.StringVar(value="6"),
            'target_fsh': tk.StringVar(value="-")
        }
        self.paut_eval_mode = tk.StringVar(value="ASME") # ASME or API
        
        # [NEW] Centralized NDT Consumable Definitions
        self.ndt_groups = {
            'PT약품': ['세척제', '침투제', '현상제', '형광침투제'],
            'MT약품': ['백색페인트', '흑색자분', '형광자분']
        }
        self.ndt_materials_all = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
        self.daily_units = ['EA', 'CAN', 'SET', 'KG', 'M', '매', 'I/D', 'P,M,I/D', 'M,I/D', 'Point', 'Meter', 'Inch', 'Dia']
        self.vehicle_boxes = []
        self.ndt_company_entries = []
        self.companies = ["원자력건설", "세안기술", "삼영검사기술", "에이치엔엘", "대한검사기술", "고려공업검사"]
        self.sites = []
        self.equipments = []
        
        # [NEW] PAUT Column Management
        self.paut_column_keys = ["selected", "Line No.", "Joint No.", "Th'k(mm)", "Start", "End", "Length(mm)", "Upper", "Lower", "Height(mm)", "Type of Flaw", "a/l", "a/t", "Evaluation", "Remarks"]

        # --- Photo Log Variables ---
        self.photo_orderer = tk.StringVar(value=self.config.get('PHOTO_ORDERER', "서울에너지공사"))
        self.photo_inspect_date = tk.StringVar(value=datetime.datetime.now().strftime("%Y-%m-%d"))
        self.photo_report_no = tk.StringVar(value=self.config.get('PHOTO_REPORT_NO', "SIT/GI-SE-NDT-2024-001"))
        self.photo_report_title = tk.StringVar(value=self.config.get('PHOTO_REPORT_TITLE', "RT 성적서 사진대장"))
        self.photo_selected_files = []
        self.photo_inspect_type = tk.StringVar(value=self.config.get('PHOTO_INSPECT_TYPE', "RT"))
        self.photo_logo_path = tk.StringVar(value=self.config.get('PHOTO_LOGO_PATH', ""))
        self.photo_cols_per_row = tk.StringVar(value=self.config.get('PHOTO_COLS_PER_ROW', "2"))
        self.photo_keep_aspect = tk.BooleanVar(value=self.config.get('PHOTO_KEEP_ASPECT', True))
        self.photo_cell_width_var = tk.StringVar(value=self.config.get('PHOTO_CELL_WIDTH', "32"))
        self.photo_cell_height_var = tk.StringVar(value=self.config.get('PHOTO_CELL_HEIGHT', "200"))
        self.photo_margin_top_var = tk.StringVar(value=self.config.get('PHOTO_MARGIN_TOP', "0.5"))
        self.photo_margin_bottom_var = tk.StringVar(value=self.config.get('PHOTO_MARGIN_BOTTOM', "0.5"))
        self.photo_margin_left_var = tk.StringVar(value=self.config.get('PHOTO_MARGIN_LEFT', "0.4"))
        self.photo_margin_right_var = tk.StringVar(value=self.config.get('PHOTO_MARGIN_RIGHT', "0.4"))
        self.photo_desc_height_var = tk.StringVar(value=self.config.get('PHOTO_DESC_HEIGHT', "25"))
        self.photo_print_scale_var = tk.StringVar(value=self.config.get('PHOTO_PRINT_SCALE', "100"))
        self.photo_align_var = tk.StringVar(value=self.config.get('PHOTO_ALIGN', "중앙 정렬"))
        self.photo_fit_width_var = tk.BooleanVar(value=self.config.get('PHOTO_FIT_WIDTH', True))
        self.photo_auto_rotate_var = tk.BooleanVar(value=self.config.get('PHOTO_AUTO_ROTATE', True))
        self.photo_width_pct_var = tk.StringVar(value=self.config.get('PHOTO_WIDTH_PCT', "100"))
        self.photo_width_pixel_adj_var = tk.StringVar(value=self.config.get('PHOTO_WIDTH_ADJ', "0"))
        self.photo_shift_x_var = tk.StringVar(value=self.config.get('PHOTO_SHIFT_X', "0"))
        self.photo_shift_y_var = tk.StringVar(value=self.config.get('PHOTO_SHIFT_Y', "0"))
        self.photo_logo_width_var = tk.StringVar(value=self.config.get('PHOTO_LOGO_WIDTH', "120"))
        self.photo_logo_x_var = tk.StringVar(value=self.config.get('PHOTO_LOGO_X', "5"))
        self.photo_logo_y_var = tk.StringVar(value=self.config.get('PHOTO_LOGO_Y', "0"))
        self.photo_output_name = tk.StringVar(value=self.config.get('PHOTO_OUTPUT_NAME', "Photo_Report"))
        self.last_photo_save_dir = self.config.get('LAST_PHOTO_SAVE_DIR', "")
        self.photo_header_map = {
            "PAUT": "REPORT OF PHASED ARRAY UT EXAMINATION (위 상 배 열 초 음 파 탐 상 검 사 보 고 서)",
            "RT": "REPORT OF RADIOGRAPHIC EXAMINATION (방 사 선 투 과 검 사 보 고 서)",
            "PT": "REPORT OF LIQUID PENETRANT EXAMINATION (침 투 탐 상 검 사 보 고 서)",
            "MT": "REPORT OF MAGNETIC PARTICLE EXAMINATION (자 분 탐 상 검 사 보 고 서)",
            "PMI": "REPORT OF POSITIVE MATERIAL IDENTIFICATION (재 질 성 분 분 석 검 사 보 고 서)",
            "UT": "REPORT OF ULTRASONIC EXAMINATION (초 음 파 탐 상 검 사 보 고 서)",
            "NDT": "REPORT OF NON-DESTRUCTIVE EXAMINATION (비 파 괴 검 사 보 고 서)",
            "기타 (직접 입력)": ""
        }
        # [NEW] Column Keys for Treeview Preview
        self.column_keys = ["_status", "selected", "No", "Date", "Dwg", "Joint", "Loc", "Ni", "Cr", "Mo", "Mn", "Grade", "Result"]
        self.rt_column_keys = ["selected", "No", "Date", "Dwg", "Joint", "Loc", "Acc", "Rej", "Deg", "D1", "D2", "D3", "D4", "D5", "D6", "D7", "D8", "D9", "D10", "D11", "D12", "D13", "D14", "D15", "Welder", "Remarks"]
        self.pt_column_keys = ["selected", "No", "Date", "Dwg", "Joint", "NPS", "Thk.", "Material", "Welder", "WType", "Result"]
        self.show_selected_only = tk.BooleanVar(value=False)
        self.auto_verify = tk.BooleanVar(value=True)
        self.extraction_mode = tk.StringVar(value="전체")
        self.sequence_filter = tk.StringVar()
        self.logo_folder_path = tk.StringVar(value=self.resource_dir)
        self.setting_vars = {}
        self.element_filters = []
        self.pmi_pane_ratio = float(self.config.get('PMI_SASH_RATIO', 0.5))
        self.item_idx_map = []
        self.rt_item_idx_map = []
        self.pt_item_idx_map = []
        self.paut_item_idx_map = []

        # --- Initialize & Load ---
        self.setup_styles()
        self.load_data()
        self.create_widgets()
        self.load_tab_config()
        self.is_ready = True
        
    def setup_styles(self):
        self.style = ttk.Style()
        # Clean modern aesthetics
        self.style.theme_use('clam')
        self.style.configure("Treeview", rowheight=28, font=("Malgun Gothic", 9))
        self.style.configure("Treeview.Heading", font=("Malgun Gothic", 9, "bold"), background="#f3f4f6")
        self.style.map("Treeview", background=[('selected', '#3b82f6')], foreground=[('selected', '#ffffff')])
        
        # Accent button style
        self.style.configure('Accent.TButton', font=('Malgun Gothic', 9, 'bold'), foreground='#ffffff', background='#3b82f6')
        self.style.map('Accent.TButton', background=[('active', '#2563eb'), ('pressed', '#1d4ed8')])

    def load_settings(self):
        if os.path.exists(self.settings_file):
            try:
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except: pass
        return {}

    def load_template_specific_config(self, template_path, mode="RT"):
        """Load JSON config linked to the Excel template file."""
        if not template_path or not os.path.exists(template_path): return
        config_path = os.path.splitext(template_path)[0] + ".json"
        if not os.path.exists(config_path): return
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                t_config = json.load(f)
            print(f"📂 [Auto-Load] Loaded specific config for {os.path.basename(template_path)}")
            for key, value in t_config.items():
                self.config[key] = value
                # Note: Final version might not use setting_vars mapping for all, 
                # but it uses config directly in many places.
                # If specific mode setup needs refresh, it will happen during run_process.
        except Exception as e:
            print(f"Error loading template config: {e}")

    def save_template_specific_config(self, mode="RT"):
        """Save current UI/config state to a template-specific JSON file."""
        if mode == "RT": template_path = self.rt_template_file_path.get()
        elif mode == "PT": template_path = self.pt_template_file_path.get()
        elif mode == "PAUT": template_path = self.paut_template_file_path.get()
        else: template_path = self.template_file_path.get()
        
        if not template_path or not os.path.exists(template_path):
            messagebox.showwarning("경고", "먼저 템플릿 파일을 선택해주세요.")
            return
            
        config_path = os.path.splitext(template_path)[0] + ".json"
        
        # Determine relevant keys to save for this template
        # 1. Logos & Print settings (Template-dependent)
        keys_to_save = []
        for ctx in ["COVER", "DATA"]:
            for prefix in ["SITCO", "SEOUL", "FOOTER", "FOOTER_PT"]:
                for suffix in ["_PATH", "_ANCHOR", "_W", "_H", "_X", "_Y"]:
                    keys_to_save.append(f"{prefix}_{mode}_{ctx}{suffix}")
            for suffix in ["_TOP", "_BOTTOM", "_LEFT", "_RIGHT"]:
                keys_to_save.append(f"MARGIN_{mode}_{ctx}{suffix}")
            keys_to_save.append(f"PRINT_SCALE_{mode}_{ctx}")
            keys_to_save.append(f"PRINT_AREA_{mode}_{ctx}")
            keys_to_save.append(f"CUSTOM_ROWS_{mode}_{ctx}")
            keys_to_save.append(f"CUSTOM_ROW_HEIGHT_{mode}_{ctx}")
            
        if mode == "RT":
            keys_to_save += ["RT_START_ROW", "RT_END_ROW"]
            keys_to_save += [k for k in self.config.keys() if k.startswith("RT_COL_") or k.startswith("RT_NAME_")]
        elif mode == "PT":
            keys_to_save += ["PT_START_ROW", "PT_END_ROW"]
            keys_to_save += [k for k in self.config.keys() if k.startswith("PT_COL_") or k.startswith("PT_NAME_")]
        elif mode == "PAUT":
            keys_to_save += ["PAUT_START_ROW", "PAUT_END_ROW"]
        elif mode == "PMI":
            keys_to_save += ["PMI_START_ROW", "PMI_DATA_END_ROW", "PMI_PRINT_END_ROW"]
            keys_to_save += [k for k in self.config.keys() if k.startswith("PMI_COL_") or k.startswith("PMI_NAME_")]

        t_data = {k: self.config[k] for k in keys_to_save if k in self.config}
        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(t_data, f, indent=4, ensure_ascii=False)
            messagebox.showinfo("저장 완료", f"템플릿 전용 설정이 저장되었습니다.\n{os.path.basename(config_path)}")
        except Exception as e:
            messagebox.showerror("저장 실패", f"설정 저장 중 오류 발생: {e}")

    def save_photo_log_config(self):
        """Save Photo Log specific layout settings to a separate JSON."""
        config_path = os.path.join(os.getcwd(), "photolog_config.json")
        self.manual_save_settings() # Capture current UI to self.config
        keys = [
            "PHOTO_MARGIN_TOP", "PHOTO_MARGIN_BOTTOM", "PHOTO_MARGIN_LEFT", "PHOTO_MARGIN_RIGHT",
            "PHOTO_PRINT_SCALE", "PHOTO_DESC_HEIGHT", "PHOTO_CELL_WIDTH", "PHOTO_CELL_HEIGHT",
            "PHOTO_COLS_PER_ROW", "PHOTO_ALIGN", "PHOTO_FIT_WIDTH", "PHOTO_KEEP_ASPECT",
            "PHOTO_AUTO_ROTATE", "PHOTO_WIDTH_PCT", "PHOTO_WIDTH_PIXEL_ADJ", 
            "PHOTO_SHIFT_X", "PHOTO_SHIFT_Y", "PHOTO_LOGO_PATH", "PHOTO_LOGO_W", "PHOTO_LOGO_X", "PHOTO_LOGO_Y"
        ]
        data = {k: self.config.get(k, "") for k in keys}
        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            print("💾 Photo Log config saved.")
            messagebox.showinfo("저장 완료", "사진대장 레이아웃 설정이 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패: {e}")

    def load_photo_log_config(self):
        """Load Photo Log specific layout settings."""
        config_path = os.path.join(os.getcwd(), "photolog_config.json")
        if not os.path.exists(config_path): return
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            for k, v in data.items():
                self.config[k] = v
                # Update UI vars if they exist
                var_name = k.lower() + "_var"
                if hasattr(self, var_name):
                    getattr(self, var_name).set(str(v))
                elif hasattr(self, k.lower()):
                    getattr(self, k.lower()).set(str(v))
            print("📂 Photo Log config loaded.")
        except: pass

    def save_settings(self):
        try:
            with open(self.settings_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"Error saving settings: {e}")

    def load_data(self):
        try:
            if not os.path.exists(self.db_path):
                self._create_empty_db()
            
            # Helper to normalize column names (remove spaces, etc.)
            def normalize_cols(df):
                df.columns = [re.sub(r'\s+', '', str(c)) for c in df.columns]
                return df

            # 1. Materials
            self.materials_df = pd.read_excel(self.db_path, sheet_name='Materials')
            self.materials_df = normalize_cols(self.materials_df)
            self.materials_df = self._sync_dataframe_schema(self.materials_df, 'Materials')
            
            # 2. Transactions
            self.transactions_df = pd.read_excel(self.db_path, sheet_name='Transactions')
            self.transactions_df = normalize_cols(self.transactions_df)
            self.transactions_df = self._sync_dataframe_schema(self.transactions_df, 'Transactions')
            
            # 3. MonthlyUsage
            try:
                self.monthly_usage_df = pd.read_excel(self.db_path, sheet_name='MonthlyUsage')
                self.monthly_usage_df = normalize_cols(self.monthly_usage_df)
                self.monthly_usage_df = self._sync_dataframe_schema(self.monthly_usage_df, 'MonthlyUsage')
            except:
                self.monthly_usage_df = pd.DataFrame(columns=['MaterialID', 'Year', 'Month', 'Site', 'Usage', 'Note', 'EntryDate'])
            
            # 4. DailyUsage
            try:
                self.daily_usage_df = pd.read_excel(self.db_path, sheet_name='DailyUsage')
                self.daily_usage_df = normalize_cols(self.daily_usage_df)
                self.daily_usage_df = self._sync_dataframe_schema(self.daily_usage_df, 'DailyUsage')
            except:
                self.daily_usage_df = pd.DataFrame(columns=['Date', 'Site', 'MaterialID', 'Usage', 'Note', 'EntryTime'])
            
            # 5. Budget
            try:
                self.budget_df = pd.read_excel(self.db_path, sheet_name='Budget')
                self.budget_df = normalize_cols(self.budget_df)
                self.budget_df = self._sync_dataframe_schema(self.budget_df, 'Budget')
            except:
                self.budget_df = pd.DataFrame(columns=['Site', 'Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit', 'Note'])
            
            # 6. Settings (Rates, Users, Vehicles, Sites)
            try:
                self.settings_df = pd.read_excel(self.db_path, sheet_name='Settings')
                self.settings_df = normalize_cols(self.settings_df)
            except:
                self.settings_df = self._create_default_settings()

            # --- Global Data Sanitization & Dtype Enforcement ---
            sanitization_map = {
                'budget_df': ['Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit'],
                'daily_usage_df': ['Usage', '검사량', '단가', '출장비', '일식', '검사비', '수량', 
                                    'RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크', 
                                    'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타'],
                'materials_df': ['가격', '원가', '수량', '재고하한'],
                'transactions_df': ['Quantity']
            }
            
            for attr, cols in sanitization_map.items():
                df = getattr(self, attr, None)
                if df is not None:
                    for col in cols:
                        if col in df.columns:
                            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0).astype(float)

            # Sync dynamic lists
            self.refresh_inquiry_filters()
            
            # Populate master lists from history
            hist_sites = self.daily_usage_df['Site'].dropna().unique().tolist() if 'Site' in self.daily_usage_df.columns else []
            trans_sites = self.transactions_df['Site'].dropna().unique().tolist() if 'Site' in self.transactions_df.columns else []
            self.sites = sorted(list(set(hist_sites + trans_sites + ["신규현장"])))
            
            if '장비명' in self.daily_usage_df.columns:
                self.equipments = sorted(self.daily_usage_df['장비명'].dropna().unique().tolist())
            if not self.equipments:
                self.equipments = ["YOKE(GY-1)", "PAUT(OmniScan X3)", "RT(XY-200)"]
            
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("데이터 오류", f"데이터를 불러오는데 실패했습니다: {e}")

    def _sync_dataframe_schema(self, df, sheet_name):
        """Ensure dataframe has all required columns for its sheet type"""
        schemas = {
            'Materials': ['MaterialID', '회사코드', '관리품번', '품목명', 'SN', '창고', '모델명', '규격', '품목군코드', '공급업체', '제조사', '제조국', '가격', '원가', '관리단위', '수량', '재고하한', 'Active'],
            'Transactions': ['Date', 'MaterialID', 'Site', 'Type', 'Quantity', 'Note', 'User', '차량번호', '주행거리', '차량점검', '차량비고'],
            'MonthlyUsage': ['MaterialID', 'Year', 'Month', 'Site', 'Usage', 'Note', 'EntryDate'],
            'DailyUsage': ['Date', 'Site', 'MaterialID', 'Usage', 'Note', 'EntryTime', 'User', '장비명', '검사량', '단가', '출장비', '일식', '검사비', '차량번호', '주행거리', '차량점검', '차량비고'],
            'Budget': ['Site', 'Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit', 'Note', 'LaborDetail', 'MaterialDetail', 'ExpenseDetail'],
            'Settings': ['Category', 'Name', 'Spec', 'Unit', 'Rate']
        }
        
        required = schemas.get(sheet_name, [])
        for col in required:
            if col not in df.columns:
                if col in ['MaterialID', 'Usage', 'Quantity', 'Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit', '가격', '원가', '수량', '재고하한']:
                    df[col] = 0.0
                elif col == 'Active':
                    df[col] = 1
                else:
                    df[col] = ""
        return df

    def _create_empty_db(self):
        with pd.ExcelWriter(self.db_path, engine='openpyxl') as writer:
            pd.DataFrame(columns=['MaterialID', '품목명', '모델명', '수량', 'Active']).to_excel(writer, sheet_name='Materials', index=False)
            pd.DataFrame(columns=['Date', 'MaterialID', 'Type', 'Quantity']).to_excel(writer, sheet_name='Transactions', index=False)
            pd.DataFrame(columns=['MaterialID', 'Year', 'Month', 'Usage']).to_excel(writer, sheet_name='MonthlyUsage', index=False)
            pd.DataFrame(columns=['Date', 'Site', 'Usage']).to_excel(writer, sheet_name='DailyUsage', index=False)
            pd.DataFrame(columns=['Site', 'Profit']).to_excel(writer, sheet_name='Budget', index=False)
            self._create_default_settings().to_excel(writer, sheet_name='Settings', index=False)

    def _create_default_settings(self):
        # Default labor, material, etc.
        data = [
            ['Labor', '이사', '', '', 55250000], ['Labor', '부장', '', '', 55250000],
            ['Material', 'PT 약품', '세척제', 'CAN', 1500], ['Material', 'PT 약품', '침투제', 'CAN', 2300]
        ]
        return pd.DataFrame(data, columns=['Category', 'Name', 'Spec', 'Unit', 'Rate'])
    def refresh_inquiry_filters(self):
        """데이터프레임으로부터 사이트, 사용자, 차량, 자재 목록을 업데이트"""
        try:
            if self.materials_df is not None:
                active_mats = self.materials_df[self.materials_df['Active'] != 0]
                self.item_name_list = sorted([str(x) for x in active_mats['품목명'].unique() if pd.notna(x)])
                self.model_list = sorted([str(x) for x in active_mats['모델명'].unique() if pd.notna(x)])
                
            if self.transactions_df is not None:
                self.sites = sorted([str(s) for s in self.transactions_df['Site'].unique() if pd.notna(s) and str(s).lower() != 'nan'])
                
            if self.daily_usage_df is not None:
                daily_sites = [str(s) for s in self.daily_usage_df['Site'].unique() if pd.notna(s) and str(s).lower() != 'nan']
                self.sites = sorted(list(set(self.sites + daily_sites)))
                
                all_users = []
                for i in range(1, 11):
                    col = 'User' if i == 1 else f'User{i}'
                    if col in self.daily_usage_df.columns:
                        all_users.extend([str(u) for u in self.daily_usage_df[col].unique() if pd.notna(u)])
                self.users = sorted([u for u in set(all_users) if u and u.lower() != 'nan'])
                
                if '차량번호' in self.daily_usage_df.columns:
                    v_list = [str(v) for v in self.daily_usage_df['차량번호'].unique() if pd.notna(v)]
                    self.vehicles = sorted([v for v in v_list if v and v.lower() != 'nan'])
                    
            if self.budget_df is not None:
                self.budget_sites = sorted([str(s) for s in self.budget_df['Site'].unique() if pd.notna(s) and str(s).lower() != 'nan'])
                
        except Exception as e:
            print(f"Error refreshing filters: {e}")

    def create_widgets(self):
        # Notebook for Tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill='both', padx=10, pady=10)
        
        # --- Material Management Tabs ---
        self.tab_stock = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_stock, text='현재 재고 현황')
        self.setup_stock_tab()
        
        self.tab_inout = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_inout, text='입출고 관리')
        self.setup_inout_tab()
        
        self.tab_daily_usage = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_daily_usage, text='현장 일일 사용량')
        self.setup_daily_usage_tab()
        
        self.tab_budget = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_budget, text='공사실행예산서')
        self.setup_budget_tab(self.tab_budget)

        # --- NDT Reporting Tabs ---
        self.tab_pmi = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_pmi, text='PMI 리포트')
        self.setup_pmi_tab(self.tab_pmi)
        
        self.tab_rt = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_rt, text='RT 리포트')
        self.setup_rt_tab(self.tab_rt)
        
        self.tab_pt = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_pt, text='PT 리포트')
        self.setup_pt_tab(self.tab_pt)
        
        self.tab_paut = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_paut, text='PAUT 리포트')
        self.setup_paut_tab(self.tab_paut)

        self.tab_photo = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_photo, text='사진대장')
        self.setup_photo_log_tab(self.tab_photo)

        # --- Settings & Utilities ---
        self.tab_import = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_import, text='데이터 관리')
        self.setup_import_tab()

        # --- Global Status Bar & Progress ---
        status_frame = tk.Frame(self.root, background="#f3f4f6", height=30)
        status_frame.pack(side='bottom', fill='x', padx=10, pady=(0, 5))
        
        self.progress = ttk.Progressbar(status_frame, orient="horizontal", length=200, mode="determinate")
        self.progress.pack(side='right', padx=10)
        
        self.status_var = tk.StringVar(value="준비 완료")
        tk.Label(status_frame, textvariable=self.status_var, background="#f3f4f6", font=("Malgun Gothic", 9)).pack(side='left', padx=10)
        
        # Internal Log for engine tracking
        self.log_text = tk.StringVar(value="")

    def log(self, message):
        print(message)
        self.status_var.set(message)
        self.root.update_idletasks()

    def setup_stock_tab(self):
        # Control Frame
        control_frame = ttk.Frame(self.tab_stock)
        control_frame.pack(fill='x', padx=5, pady=5)
        
        # Row 1: Action Buttons
        action_row = ttk.Frame(control_frame)
        action_row.pack(fill='x', side='top', pady=(0, 5))
        
        ttk.Button(action_row, text="🔄 새로고침", command=self.update_stock_view).pack(side='left', padx=2)
        ttk.Button(action_row, text="⚠️ 재고부족 알림", command=self.show_low_stock).pack(side='left', padx=2)
        ttk.Button(action_row, text="🗑️ 품목 삭제", command=self.delete_selected_material).pack(side='left', padx=2)
        ttk.Button(action_row, text="📝 정보 수정", command=self.open_edit_material_dialog).pack(side='left', padx=2)
        ttk.Button(action_row, text="📊 엑셀 내보내기", command=self.export_stock_to_excel).pack(side='left', padx=2)
        ttk.Button(action_row, text="✅ 전체 선택", command=self.select_all_stock).pack(side='left', padx=2)
        
        # Row 2: Search and Filter
        filter_row = ttk.Frame(control_frame)
        filter_row.pack(fill='x', side='top')
        filter_frame = ttk.LabelFrame(filter_row, text=" 상세 검색 필터 ")
        filter_frame.pack(fill='x', padx=5, pady=2)
        
        ttk.Label(filter_frame, text="회사:").grid(row=0, column=0, padx=2, pady=2, sticky='e')
        self.cb_filter_co = ttk.Combobox(filter_frame, width=12); self.cb_filter_co.grid(row=0, column=1, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="분류:").grid(row=0, column=2, padx=2, pady=2, sticky='e')
        self.cb_filter_class = ttk.Combobox(filter_frame, width=12); self.cb_filter_class.grid(row=0, column=3, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="제조사:").grid(row=0, column=4, padx=2, pady=2, sticky='e')
        self.cb_filter_mfr = ttk.Combobox(filter_frame, width=12); self.cb_filter_mfr.grid(row=0, column=5, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="품목명:").grid(row=0, column=6, padx=2, pady=2, sticky='e')
        self.cb_filter_name = ttk.Combobox(filter_frame, width=20); self.cb_filter_name.grid(row=0, column=7, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="S/N:").grid(row=1, column=0, padx=2, pady=2, sticky='e')
        self.cb_filter_sn = ttk.Combobox(filter_frame, width=12); self.cb_filter_sn.grid(row=1, column=1, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="모델명:").grid(row=1, column=2, padx=2, pady=2, sticky='e')
        self.cb_filter_model = ttk.Combobox(filter_frame, width=12); self.cb_filter_model.grid(row=1, column=3, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="관리품목:").grid(row=1, column=4, padx=2, pady=2, sticky='e')
        self.cb_filter_eq = ttk.Combobox(filter_frame, width=20)
        self.cb_filter_eq.grid(row=1, column=5, padx=2, pady=2)
        
        stock_filters = [
            self.cb_filter_co, self.cb_filter_class, self.cb_filter_mfr,
            self.cb_filter_name, self.cb_filter_sn, self.cb_filter_model,
            self.cb_filter_eq
        ]

        for combo in stock_filters:
            combo.bind('<<ComboboxSelected>>', lambda e: self.update_stock_view())
        
        ttk.Label(filter_frame, text="검색어:").grid(row=1, column=6, padx=2, pady=2, sticky='e')
        self.search_var = tk.StringVar()
        self.search_var.trace_add('write', lambda *args: self.update_stock_view())
        self.search_entry = ttk.Entry(filter_frame, textvariable=self.search_var, width=20)
        self.search_entry.grid(row=1, column=7, padx=2, pady=2)
        
        # Reset Filters Button
        btn_reset = ttk.Button(filter_frame, text="♻️ 필터 초기화", command=self.reset_stock_filters)
        btn_reset.grid(row=1, column=8, padx=10, pady=2)
        
        # Treeview for Stock with Scrollbars
        tree_frame = ttk.Frame(self.tab_stock)
        tree_frame.pack(expand=True, fill='both', padx=5, pady=5)
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        # 18 Columns matching the high-fidelity original system
        columns = ('MaterialID', 'AdminCode', 'Category', 'ItemName', 'SN', 'Warehouse', 
                   'Model', 'Spec', 'GroupCode', 'Supplier', 'Maker', 'MakeDate', 
                   'Price', 'Location', 'Unit', 'BaseQty', 'CurrentStock', 'Limit')
        
        self.stock_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', 
                                      yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        vsb.config(command=self.stock_tree.yview)
        hsb.config(command=self.stock_tree.xview)
        
        # Heading Map (Korean)
        col_map = {
            'MaterialID': 'ID', 'AdminCode': '관리사코드', 'Category': '관리품목', 'ItemName': '품목명',
            'SN': 'SN', 'Warehouse': '창고', 'Model': '모델명', 'Spec': '규격',
            'GroupCode': '품목군코드', 'Supplier': '공급업체', 'Maker': '제조사', 'MakeDate': '제조일',
            'Price': '가격', 'Location': '위치', 'Unit': '관리단위', 'BaseQty': '기초수량',
            'CurrentStock': '현재재고', 'Limit': '재고기한'
        }
        
        for col in columns:
            self.stock_tree.heading(col, text=col_map.get(col, col))
            # Set specific widths for better readability
            w = 120 if col in ['ItemName', 'Model', 'Supplier'] else 80
            self.stock_tree.column(col, width=w, anchor='center')
        
        self.stock_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def select_all_stock(self):
        """Select all items in the stock treeview"""
        all_items = self.stock_tree.get_children()
        self.stock_tree.selection_set(all_items)

    def show_low_stock(self):
        """Show items with low stock (less than their specific reorder point)"""
        low_stock_items = []
        for _, mat in self.materials_df.iterrows():
            if mat.get('Active', 1) == 0:
                continue
                
            current = self.calculate_current_stock(mat['MaterialID'])
            reorder_point = mat.get('재고하한', 10)
            if pd.isna(reorder_point) or reorder_point <= 0:
                reorder_point = 10 # Default fallback
                
            if pd.notna(current) and current < reorder_point:
                low_stock_items.append((mat.get('품목명', ''), current, reorder_point))
        
        if not low_stock_items:
            messagebox.showinfo("재고 알림", "수량이 재고하한 미만인 항목이 없습니다.")
        else:
            msg = "다음 항목들의 재고가 부족합니다:\n\n"
            for item, current, reorder in low_stock_items:
                msg += f"• {item}: 현재 {current:g} (재고하한: {reorder:g})\n"
            messagebox.showwarning("재고 부족", msg)



    def setup_import_tab(self):
        # Implementation for Data Management tab
        ttk.Label(self.tab_import, text="데이터 관리 및 엑셀 업로드 센터", font=("Malgun Gothic", 12, "bold")).pack(pady=20)
        ttk.Button(self.tab_import, text="마스터 데이터 엑셀 업로드", command=self.import_master_data).pack(pady=10)
        ttk.Button(self.tab_import, text="전체 데이터 백업", command=self.backup_all_data).pack(pady=10)

    def setup_inout_tab(self):
        container = tk.Frame(self.tab_inout, background="#f9fafb")
        container.pack(fill='both', expand=True)
        
        # Dual Pane
        paned = tk.PanedWindow(container, orient='horizontal', background="#d1d5db", sashwidth=6)
        paned.pack(fill='both', expand=True)
        
        # [LEFT] Registration Form
        left_f = ttk.Frame(paned, padding=10)
        paned.add(left_f, width=420)
        
        ttk.Label(left_f, text="📥 입출고 등록", font=("Malgun Gothic", 12, "bold")).pack(pady=(0, 10))
        
        form = ttk.Frame(left_f)
        form.pack(fill='x')
        
        fields = [("날짜:", "Date"), ("품목:", "Material"), ("현장:", "Site"), ("구분:", "Type"), ("수량:", "Qty"), ("비고:", "Note")]
        self.inout_vars = {}
        for i, (lbl, key) in enumerate(fields):
            ttk.Label(form, text=lbl).grid(row=i, column=0, padx=5, pady=5, sticky='e')
            
            f_inner = ttk.Frame(form)
            f_inner.grid(row=i, column=1, padx=5, pady=5, sticky='w')
            
            if key == "Date":
                var = DateEntry(f_inner, width=15, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
            elif key == "Material":
                var = ttk.Combobox(f_inner, width=25, values=self.item_name_list)
                tk.Button(f_inner, text="🔍", font=('Arial', 8), bd=0, bg='white', cursor='hand2', 
                          command=lambda: self.open_material_search_dialog(target_form='inout')).pack(side='right', padx=2)
            elif key == "Site":
                var = ttk.Combobox(f_inner, width=25, values=self.sites)
                tk.Button(f_inner, text="🔍", font=('Arial', 8), bd=0, bg='white', cursor='hand2', 
                          command=lambda: self.open_list_management_dialog('sites', target_cb=var)).pack(side='right', padx=2)
            elif key == "Type":
                var = ttk.Combobox(f_inner, width=25)
                var['values'] = ["입고", "출고", "반납", "폐기", "수정"]
                var.set("입고")
            else:
                var = ttk.Entry(f_inner, width=28)
            
            var.pack(side='left')
            self.inout_vars[key] = var

        # Button Row for Registration
        btn_f = ttk.Frame(left_f)
        btn_f.pack(pady=20)
        ttk.Button(btn_f, text="등록하기", style='Accent.TButton', command=self.register_transaction, width=15).pack(side='left', padx=5)
        ttk.Button(btn_f, text="초기화", command=self.clear_inout_form, width=10).pack(side='left', padx=5)

        # [RIGHT] History & Search
        right_f = ttk.Frame(paned, padding=10)
        paned.add(right_f, stretch="always")
        
        # Search & Filter Frame
        search_f = ttk.Frame(right_f)
        search_f.pack(fill='x', pady=(0, 10))
        
        ttk.Label(search_f, text="🔍 검색:").pack(side='left', padx=5)
        self.inout_search_var = tk.StringVar()
        self.inout_search_var.trace_add("write", self.on_inout_search_change)
        search_ent = ttk.Entry(search_f, textvariable=self.inout_search_var, width=30)
        search_ent.pack(side='left', padx=5)
        
        ttk.Button(search_f, text="새로고침", command=self.update_transaction_view, width=10).pack(side='left', padx=5)
        ttk.Button(search_f, text="엑셀 저장", command=self.export_inout_history, width=10).pack(side='right', padx=5)
        ttk.Button(search_f, text="선택 삭제", command=self.delete_transaction_entry, width=10).pack(side='right', padx=5)

        # Treeview
        cols = ("날짜", "현장", "구분", "품목", "수량", "비고")
        self.inout_tree = ttk.Treeview(right_f, columns=cols, show='headings', height=15)
        for c in cols: 
            anchor = 'center' if c != "비고" else 'w'
            width = 150 if c in ["날짜", "품목"] else 80
            self.inout_tree.heading(c, text=c)
            self.inout_tree.column(c, width=width, anchor=anchor)
        
        # Scrollbar for Treeview
        vsb = ttk.Scrollbar(right_f, orient="vertical", command=self.inout_tree.yview)
        self.inout_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side='right', fill='y')
        self.inout_tree.pack(fill='both', expand=True)
        
        # Summary Label
        self.inout_summary_var = tk.StringVar(value="총 0건의 내역")
        ttk.Label(right_f, textvariable=self.inout_summary_var, font=("Malgun Gothic", 9, "italic")).pack(anchor='e', pady=5)

        # Bind Enter keys for form navigation
        self.inout_vars['Date'].bind('<Return>', lambda e: self.inout_vars['Material'].focus_set())
        self.inout_vars['Material'].bind('<Return>', lambda e: self.inout_vars['Site'].focus_set())
        self.inout_vars['Site'].bind('<Return>', lambda e: self.inout_vars['Type'].focus_set())
        self.inout_vars['Type'].bind('<Return>', lambda e: self.inout_vars['Qty'].focus_set())
        self.inout_vars['Qty'].bind('<Return>', lambda e: self.inout_vars['Note'].focus_set())
        self.inout_vars['Note'].bind('<Return>', lambda e: self.register_transaction())

    def setup_daily_usage_tab(self):
        """Setup the daily usage entry tab"""
        # Create PanedWindow for resizable frames
        self.daily_usage_paned = ttk.Panedwindow(self.tab_daily_usage, orient='vertical')
        self.daily_usage_paned.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Save sash position on adjustment
        self.daily_usage_paned.bind("<ButtonRelease-1>", lambda e: self.save_tab_config())
        
        entry_frame = ttk.LabelFrame(self.daily_usage_paned, text="작업일보 및 일일 사용량 기입")
        self.daily_usage_paned.add(entry_frame, weight=1)
        
        # Header area
        header_container = ttk.Frame(entry_frame)
        header_container.pack(fill='x', padx=2, pady=1)
        
        row1 = ttk.Frame(header_container)
        row1.pack(fill='x', pady=1)
        row2 = ttk.Frame(header_container)
        row2.pack(fill='x', pady=1)
        
        # Row 1: Primary Actions
        self.btn_daily_save = ttk.Button(row1, text="💾 저장", command=self.add_daily_usage_entry, width=8)
        self.btn_daily_save.pack(side='left', padx=2)
        
        ttk.Button(row1, text="🧹 초기화", command=lambda: self.clear_daily_usage_form_all(keep_date=True), width=10).pack(side='left', padx=2)
        
        btn_ndt_map = ttk.Button(row1, text="🧪 NDT 품목 매핑", command=self.open_ndt_product_map_dialog)
        btn_ndt_map.pack(side='left', padx=5)

        btn_sync = ttk.Button(row1, text="🔄 작업자 일괄 적용", command=self.sync_worker_times, width=20)
        btn_sync.pack(side='left', padx=5)

        # Row 2: Tool Actions
        btn_add_vehicle = ttk.Button(row2, text="🚗 차량점검", command=lambda: messagebox.showinfo("알림", "차량 점검 위젯은 리포트 생성 시 자동 포함됩니다."))
        btn_add_vehicle.pack(side='right', padx=2)

        # Canvas for scrollable form
        canvas_parent = ttk.Frame(entry_frame)
        canvas_parent.pack(fill='both', expand=True, padx=2, pady=1)
        
        self.entry_canvas = tk.Canvas(canvas_parent, highlightthickness=0)
        entry_vsb = ttk.Scrollbar(canvas_parent, orient="vertical", command=self.entry_canvas.yview)
        entry_vsb.pack(side='right', fill='y')
        self.entry_canvas.configure(yscrollcommand=entry_vsb.set)
        self.entry_canvas.pack(side='left', fill='both', expand=True)
        
        self.entry_inner_frame = ttk.Frame(self.entry_canvas)
        self.entry_canvas_window = self.entry_canvas.create_window((0, 0), window=self.entry_inner_frame, anchor='nw')
        
        def _on_entry_config(e):
            target_w = max(1100, e.width)
            self.entry_canvas.itemconfig(self.entry_canvas_window, width=target_w)
            self.entry_canvas.configure(scrollregion=self.entry_canvas.bbox("all"))
        
        self.entry_inner_frame.bind("<Configure>", lambda e: self.entry_canvas.configure(scrollregion=self.entry_canvas.bbox("all")))
        self.entry_canvas.bind("<Configure>", _on_entry_config)
        
        # Form Panels
        self.master_form_panel = ttk.LabelFrame(self.entry_inner_frame, text="작업일보 현장 기록 (검사 및 사용량)")
        self.master_form_panel.pack(fill='x', padx=5, pady=5)
        
        # Split into Left (Form) and Right (Workers)
        split_row = ttk.Frame(self.master_form_panel)
        split_row.pack(fill='x', expand=True)
        
        form_content = ttk.Frame(split_row, padding=10)
        form_content.pack(side='left', anchor='nw')
        
        worker_section = ttk.LabelFrame(split_row, text="작업자 및 근무시간 기록")
        worker_section.pack(side='left', fill='both', expand=True, padx=5, pady=5)
        
        # Add management buttons for core fields
        def add_mgr(parent, key, target_cb=None):
            tk.Button(parent, text="⚙️", font=('Arial', 8), bd=0, bg="#f3f4f6", fg='blue', cursor='hand2',
                      command=lambda: self.open_list_management_dialog(key, target_cb=target_cb)).pack(side='left', padx=2)

        # Basic Form Fields
        # Use Grid for 2-column layout as in V13
        # Row 0: Company, Site
        ttk.Label(form_content, text="업체명:").grid(row=0, column=0, padx=(5, 0), pady=1, sticky='e')
        co_f = ttk.Frame(form_content); co_f.grid(row=0, column=1, padx=(2, 10), pady=1, sticky='w')
        self.cb_daily_company = ttk.Combobox(co_f, width=12, values=self.companies); self.cb_daily_company.pack(side='left')
        add_mgr(co_f, 'companies', self.cb_daily_company)
        
        ttk.Label(form_content, text="현장명:").grid(row=0, column=2, padx=(5, 0), pady=1, sticky='e')
        site_f = ttk.Frame(form_content); site_f.grid(row=0, column=3, padx=(2, 5), pady=1, sticky='w')
        self.cb_daily_site = ttk.Combobox(site_f, width=12, values=self.sites); self.cb_daily_site.pack(side='left')
        add_mgr(site_f, 'sites', self.cb_daily_site)
        
        # Row 1: Date, Equipment
        ttk.Label(form_content, text="날짜:").grid(row=1, column=0, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_date = DateEntry(form_content, width=15, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly')
        self.ent_daily_date.grid(row=1, column=1, padx=(2, 10), pady=1, sticky='w')
        
        ttk.Label(form_content, text="장비명:").grid(row=1, column=2, padx=(5, 0), pady=1, sticky='e')
        eq_f = ttk.Frame(form_content); eq_f.grid(row=1, column=3, padx=(2, 5), pady=1, sticky='w')
        self.cb_daily_equip = ttk.Combobox(eq_f, width=12, values=self.equipments); self.cb_daily_equip.pack(side='left')
        tk.Button(self.cb_daily_equip, text="🔍", font=('Arial', 8), bd=0, bg='white', cursor='hand2', 
                  command=self.open_equipment_search_dialog).place(relx=1.0, x=-2, rely=0.5, anchor='e', width=18, height=18)
        add_mgr(eq_f, 'equipments', self.cb_daily_equip)
        
        # Row 2: Material (Full width)
        ttk.Label(form_content, text="품목명:").grid(row=2, column=0, padx=(5, 0), pady=1, sticky='e')
        mat_f = ttk.Frame(form_content); mat_f.grid(row=2, column=1, columnspan=3, padx=(2, 5), pady=1, sticky='w')
        self.cb_daily_material = ttk.Combobox(mat_f, width=35, values=self.item_name_list); self.cb_daily_material.pack(side='left')
        tk.Button(self.cb_daily_material, text="🔍", font=('Arial', 8), bd=0, bg='white', cursor='hand2', 
                  command=lambda: self.open_material_search_dialog(target_form='daily_usage')).place(relx=1.0, x=-2, rely=0.5, anchor='e', width=18, height=18)
        add_mgr(mat_f, 'materials', self.cb_daily_material)
        
        # Row 3: Method, Inspection Item
        ttk.Label(form_content, text="방법:").grid(row=3, column=0, padx=(5, 0), pady=1, sticky='e')
        self.cb_daily_test_method = ttk.Combobox(form_content, width=15, values=['RT', 'PAUT', 'UT', 'MT', 'PT', 'ETC'])
        self.cb_daily_test_method.grid(row=3, column=1, padx=(2, 10), pady=1, sticky='w')
        
        ttk.Label(form_content, text="검사품명:").grid(row=3, column=2, padx=(5, 0), pady=1, sticky='e')
        insp_f = ttk.Frame(form_content); insp_f.grid(row=3, column=3, padx=(2, 5), pady=1, sticky='w')
        self.ent_daily_inspection_item = ttk.Entry(insp_f, width=15); self.ent_daily_inspection_item.pack(side='left')
        self.ent_daily_inspection_item.insert(0, "Piping")
        add_mgr(insp_f, 'test_items', self.ent_daily_inspection_item)
        
        # Row 4: Qty, Unit
        ttk.Label(form_content, text="수량:").grid(row=4, column=0, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_test_amount = ttk.Entry(form_content, width=15); self.ent_daily_test_amount.grid(row=4, column=1, padx=(2, 10), pady=1, sticky='w')
        self.ent_daily_test_amount.bind('<KeyRelease>', self.update_daily_test_fee_calc)
        
        ttk.Label(form_content, text="단위:").grid(row=4, column=2, padx=(5, 0), pady=1, sticky='e')
        unit_f = ttk.Frame(form_content); unit_f.grid(row=4, column=3, padx=(2, 5), pady=1, sticky='w')
        self.cb_daily_unit = ttk.Combobox(unit_f, width=12, values=self.daily_units); self.cb_daily_unit.pack(side='left'); self.cb_daily_unit.set('매')
        add_mgr(unit_f, 'daily_units', self.cb_daily_unit)
        
        # Row 5: Price, Travel
        ttk.Label(form_content, text="단가:").grid(row=5, column=0, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_unit_price = ttk.Entry(form_content, width=15); self.ent_daily_unit_price.grid(row=5, column=1, padx=(2, 10), pady=1, sticky='w')
        self.ent_daily_unit_price.bind('<KeyRelease>', self.update_daily_test_fee_calc)
        
        ttk.Label(form_content, text="출장비:").grid(row=5, column=2, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_travel_cost = ttk.Entry(form_content, width=15); self.ent_daily_travel_cost.grid(row=5, column=3, padx=(2, 5), pady=1, sticky='w')
        self.ent_daily_travel_cost.insert(0, "0"); self.ent_daily_travel_cost.bind('<KeyRelease>', self.update_daily_test_fee_calc)
        
        # Row 6: Applied Code, Report No
        ttk.Label(form_content, text="적용코드:").grid(row=6, column=0, padx=(5, 0), pady=1, sticky='e')
        app_f = ttk.Frame(form_content); app_f.grid(row=6, column=1, padx=(2, 10), pady=1, sticky='w')
        self.ent_daily_applied_code = ttk.Entry(app_f, width=12); self.ent_daily_applied_code.pack(side='left'); self.ent_daily_applied_code.insert(0, "KS")
        add_mgr(app_f, 'applied_codes', self.ent_daily_applied_code)
        
        ttk.Label(form_content, text="성적서번호:").grid(row=6, column=2, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_report_no = ttk.Entry(form_content, width=18); self.ent_daily_report_no.grid(row=6, column=3, padx=(2, 5), pady=1, sticky='w')
        
        # Row 7: Note, Meal
        ttk.Label(form_content, text="비고:").grid(row=7, column=0, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_note = ttk.Entry(form_content, width=15); self.ent_daily_note.grid(row=7, column=1, padx=(2, 10), pady=1, sticky='w')
        
        ttk.Label(form_content, text="일식:").grid(row=7, column=2, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_meal_cost = ttk.Entry(form_content, width=15); self.ent_daily_meal_cost.grid(row=7, column=3, padx=(2, 5), pady=1, sticky='w')
        self.ent_daily_meal_cost.insert(0, "0"); self.ent_daily_meal_cost.bind('<KeyRelease>', self.update_daily_test_fee_calc)
        
        # Row 8: Fee
        ttk.Label(form_content, text="검사비:").grid(row=8, column=0, padx=(5, 0), pady=1, sticky='e')
        self.ent_daily_test_fee = ttk.Entry(form_content, width=15); self.ent_daily_test_fee.grid(row=8, column=1, padx=(2, 10), pady=1, sticky='w')
        
        # Focus transitions
        self.ent_daily_inspection_item.bind('<Return>', lambda e: self.ent_daily_test_amount.focus_set())
        self.ent_daily_test_amount.bind('<Return>', lambda e: self.cb_daily_unit.focus_set())
        self.cb_daily_unit.bind('<Return>', lambda e: self.ent_daily_unit_price.focus_set())
        self.ent_daily_unit_price.bind('<Return>', lambda e: self.ent_daily_applied_code.focus_set())
        self.ent_daily_applied_code.bind('<Return>', lambda e: self.ent_daily_report_no.focus_set())
        self.ent_daily_report_no.bind('<Return>', lambda e: self.ent_daily_note.focus_set())
        self.ent_daily_note.bind('<Return>', lambda e: self.ent_daily_meal_cost.focus_set())
        self.ent_daily_meal_cost.bind('<Return>', lambda e: self.ent_daily_test_fee.focus_set())

        # --- Worker Sections (Already inside split_row) ---
        workers_container = ttk.Frame(worker_section, padding=5)
        workers_container.pack(fill='both', expand=True)
        
        for c in range(2): workers_container.grid_columnconfigure(c, weight=1)
        
        self.worker_groups = []
        def setup_worker_group(idx, r, c):
            f = ttk.LabelFrame(workers_container, text=f"작업자 {idx}")
            f.grid(row=r, column=c, padx=2, pady=1, sticky='ew')
            group = WorkerDataGroup(f, worker_index=idx, users_list=self.users, enable_autocomplete=True, time_list=self.worktimes)
            group.pack(fill='x', expand=True, padx=1, pady=1)
            self.worker_groups.append(group)
            setattr(self, f'worker_group{idx}', group)
            # Legacy mapping for compatibility
            if idx == 1: self.cb_daily_user = group.composite; self.ent_worktime1 = group.ent_worktime; self.ent_ot1 = group.ent_ot

        for i in range(1, 6): setup_worker_group(i, i-1, 0)
        for i in range(6, 11): setup_worker_group(i, i-6, 1)

        # --- NDT Consumables Section ---
        ndt_section = ttk.LabelFrame(self.entry_inner_frame, text="NDT 약품 사용량 (업체별)")
        ndt_section.pack(fill='x', padx=5, pady=5)
        
        btn_row = ttk.Frame(ndt_section)
        btn_row.pack(fill='x', padx=5)
        ttk.Button(btn_row, text="+ 업체 추가", command=self.add_ndt_company_section, width=12).pack(side='left', padx=2)
        ttk.Button(btn_row, text="- 삭제", command=self.remove_last_ndt_company, width=10).pack(side='left', padx=2)

        self.ndt_company_container = ttk.Frame(ndt_section, padding=5)
        self.ndt_company_container.pack(fill='x')
        self.add_ndt_company_section() # Initial one

        # --- RTK Quality Defect Section ---
        rtk_section = ttk.LabelFrame(self.entry_inner_frame, text="RTK 품질 결함 (매)")
        rtk_section.pack(fill='x', padx=5, pady=5)
        
        rtk_fields = [
            ("센터미스:", "센터미스"), ("농도:", "농도"), ("마킹미스:", "마킹미스"),
            ("필름마크:", "필름마크"), ("취급부주의:", "취급부주의"), ("고객불만:", "고객불만"),
            ("기타:", "기타")
        ]
        self.rtk_vars = {}
        rtk_grid = ttk.Frame(rtk_section, padding=5)
        rtk_grid.pack(fill='x')
        
        for idx, (lbl, key) in enumerate(rtk_fields):
            r, c = divmod(idx, 4)
            ttk.Label(rtk_grid, text=lbl).grid(row=r, column=c*2, padx=5, pady=2, sticky='e')
            var = ttk.Entry(rtk_grid, width=8)
            var.grid(row=r, column=c*2+1, padx=5, pady=2, sticky='w')
            var.insert(0, "0")
            self.rtk_vars[key] = var
        
        self.lbl_rtk_total = ttk.Label(rtk_grid, text="총계: 0매", font=("Malgun Gothic", 9, "bold"))
        self.lbl_rtk_total.grid(row=1, column=6, columnspan=2, padx=20)
        
        for v in self.rtk_vars.values():
            v.bind('<KeyRelease>', lambda e: self.update_rtk_total())

        # --- History Inquiry Section (RESTORED) ---
        history_frame = ttk.LabelFrame(self.daily_usage_paned, text="작업일보 및 일일 사용현황 조회")
        self.daily_usage_paned.add(history_frame, weight=1)
        self.setup_daily_usage_history_view(history_frame)

    def update_rtk_total(self, event=None):
        total = 0
        for v in self.rtk_vars.values():
            try: total += int(v.get())
            except: pass
        self.lbl_rtk_total.config(text=f"총계: {total}매")

    def delete_daily_usage_entry(self):
        selection = self.daily_tree.selection()
        if not selection:
            messagebox.showwarning("선택 오류", "삭제할 항목을 선택해주세요.")
            return
        if not messagebox.askyesno("삭제 확인", "선택한 현장 사용 내역을 삭제하시겠습니까?"): return
        try:
            for item in selection:
                vals = self.daily_tree.item(item, 'values')
                # Date(0), Site(1), Item(2), Qty(3), Unit(4), Price(5), Fee(6), User(7)
                date_str, site, item_name = vals[0], vals[1], vals[2]
                mask = (pd.to_datetime(self.daily_usage_df['Date']).dt.strftime('%Y-%m-%d') == date_str) & \
                       (self.daily_usage_df['Site'] == site) & \
                       (self.daily_usage_df['검사품명'] == item_name)
                self.daily_usage_df = self.daily_usage_df[~mask]
            self.save_data()
            self.update_daily_usage_view()
            messagebox.showinfo("성공", "삭제되었습니다.")
        except Exception as e: messagebox.showerror("오류", f"삭제 실패: {e}")

    def setup_daily_usage_history_view(self, parent):
        """Setup the history tree and filters for daily usage"""
        # Filter Panel (3-row layout as requested)
        filter_panel = ttk.Frame(parent, padding=5)
        filter_panel.pack(fill='x')
        
        # Filter Row 1: Date Range
        row1 = ttk.Frame(filter_panel); row1.pack(fill='x', pady=1)
        ttk.Label(row1, text="시작일:").pack(side='left', padx=2)
        self.ent_daily_filter_start = DateEntry(row1, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        self.ent_daily_filter_start.pack(side='left', padx=2)
        
        ttk.Label(row1, text="종료일:").pack(side='left', padx=5)
        self.ent_daily_filter_end = DateEntry(row1, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        self.ent_daily_filter_end.pack(side='left', padx=2)
        
        # Filter Row 2: Comboboxes
        row2 = ttk.Frame(filter_panel); row2.pack(fill='x', pady=1)
        ttk.Label(row2, text="현장:").pack(side='left', padx=2)
        self.cb_daily_filter_site = ttk.Combobox(row2, width=15, values=['전체'] + self.sites); self.cb_daily_filter_site.pack(side='left', padx=2); self.cb_daily_filter_site.set('전체')
        
        ttk.Label(row2, text="작업자:").pack(side='left', padx=5)
        self.cb_daily_filter_user = ttk.Combobox(row2, width=15, values=['전체'] + self.users); self.cb_daily_filter_user.pack(side='left', padx=2); self.cb_daily_filter_user.set('전체')

        # Filter Row 3: Search & Actions
        row3 = ttk.Frame(filter_panel); row3.pack(fill='x', pady=1)
        ttk.Label(row3, text="🔍 검색:").pack(side='left', padx=2)
        self.daily_search_var = tk.StringVar()
        self.daily_search_var.trace_add("write", lambda *a: self.update_daily_usage_view())
        ttk.Entry(row3, textvariable=self.daily_search_var, width=20).pack(side='left', padx=2)

        ttk.Button(row3, text="🔍 조회", command=self.update_daily_usage_view, width=10).pack(side='left', padx=10)
        ttk.Button(row3, text="📊 엑셀 저장", command=self.export_daily_report_to_excel, width=12).pack(side='left', padx=2)
        ttk.Button(row3, text="선택 삭제", command=self.delete_daily_usage_entry, width=10).pack(side='right', padx=5)
        
        # Treeview
        cols = ("날짜", "현장명", "검사항목/품명", "수량", "단위", "단가", "검사비", "검사원")
        self.daily_tree = ttk.Treeview(parent, columns=cols, show='headings', height=10)
        for c in cols: 
            width = 120 if c in ["날짜", "현장", "검사품명"] else 80
            self.daily_tree.heading(c, text=c); self.daily_tree.column(c, width=width, anchor='center')
        
        vsb = ttk.Scrollbar(parent, orient="vertical", command=self.daily_tree.yview)
        self.daily_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side='right', fill='y')
        self.daily_tree.pack(fill='both', expand=True, padx=2, pady=2)
        
        # Summary Area
        summary_f = ttk.Frame(parent)
        summary_f.pack(fill='x', padx=5, pady=2)
        self.daily_summary_var = tk.StringVar(value="조회된 내역: 0건 | 총 검사비: 0원")
        ttk.Label(summary_f, textvariable=self.daily_summary_var, font=("Malgun Gothic", 9, "bold")).pack(side='right')

    def register_transaction(self):
        """Record an IN or OUT transaction using inout_vars"""
        try:
            mat_display = self.inout_vars['Material'].get().strip()
            if not mat_display:
                messagebox.showwarning("입력 오류", "자재를 선택해주세요.")
                return
            
            mat_id = self._get_id_from_display(mat_display)
            if not mat_id:
                if messagebox.askyesno("신규 자재", f"'{mat_display}'은(는) 등록되지 않은 자재입니다. 신규 등록하시겠습니까?"):
                    mat_id = self.register_new_material(mat_display)
                else: return

            trans_type = self.inout_vars['Type'].get()
            try:
                qty = float(self.inout_vars['Qty'].get())
            except ValueError:
                messagebox.showwarning("입력 오류", "수량을 숫자로 입력해주세요.")
                return

            site = self.inout_vars['Site'].get().strip()
            note = self.inout_vars['Note'].get().strip()
            
            # Date from DateEntry
            date_val = self.inout_vars['Date'].get_date()

            if trans_type in ['출고', '폐기']: qty = -abs(qty)
            else: qty = abs(qty)

            new_row = {
                'Date': date_val,
                'MaterialID': mat_id,
                'Type': trans_type,
                'Quantity': qty,
                'Site': site,
                'Note': note
            }
            self.transactions_df = pd.concat([self.transactions_df, pd.DataFrame([new_row])], ignore_index=True)
            self.save_data()
            self.update_transaction_view()
            self.update_stock_view()
            
            self.inout_vars['Qty'].delete(0, tk.END)
            self.inout_vars['Note'].delete(0, tk.END)
            messagebox.showinfo("성공", f"{trans_type} 기록이 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"기록 저장 중 오류 발생: {e}")

    def on_material_selected(self, event=None):
        display = self.inout_vars['Material'].get().strip()
        mat_id = self._get_id_from_display(display)
        if mat_id:
            mat_info = self.materials_df[self.materials_df['MaterialID'] == mat_id]
            if not mat_info.empty:
                # If there's a warehouse or unit we want to auto-fill, do it here
                pass

    def delete_transaction_entry(self):
        selection = self.inout_tree.selection()
        if not selection:
            messagebox.showwarning("선택 오류", "삭제할 항목을 선택해주세요.")
            return
        if not messagebox.askyesno("삭제 확인", "선택한 내역을 삭제하시겠습니까?"): return
        try:
            for item in selection:
                values = self.inout_tree.item(item, 'values')
                date_str, site, t_type, mat_disp, qty = values[0], values[1], values[2], values[3], float(values[4])
                mat_id = self._get_id_from_display(mat_disp)
                mask = (pd.to_datetime(self.transactions_df['Date']).dt.strftime('%Y-%m-%d %H:%M:%S') == date_str) & \
                       (self.transactions_df['MaterialID'] == mat_id) & (self.transactions_df['Quantity'] == qty)
                self.transactions_df = self.transactions_df[~mask]
            self.save_data()
            self.update_transaction_view()
            self.update_stock_view()
            messagebox.showinfo("성공", "삭제되었습니다.")
        except Exception as e: messagebox.showerror("오류", f"삭제 실패: {e}")

    def on_inout_search_change(self, *args):
        self.update_transaction_view()

    def clear_inout_form(self):
        """Reset the inbound/outbound registration form."""
        self.inout_vars['Date'].set_date(datetime.date.today())
        self.inout_vars['Material'].set('')
        self.inout_vars['Site'].set('')
        self.inout_vars['Type'].set('입고')
        self.inout_vars['Qty'].delete(0, tk.END)
        self.inout_vars['Note'].delete(0, tk.END)

    def export_inout_history(self):
        """Export current filtered transaction history to Excel."""
        if self.transactions_df.empty:
            messagebox.showwarning("데이터 없음", "내보낼 데이터가 없습니다.")
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"입출고내역_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not file_path: return
        
        try:
            # Apply current search filter to the exported data
            q = self.inout_search_var.get().strip().lower()
            df_export = self.transactions_df.copy()
            
            if q:
                # Add material name for filtering
                df_export['품목명'] = df_export['MaterialID'].apply(self.get_material_display_name)
                mask = df_export.apply(lambda r: q in str(r['품목명']).lower() or q in str(r.get('Site','')).lower() or q in str(r.get('Type','')).lower(), axis=1)
                df_export = df_export[mask]
                df_export.drop(columns=['품목명'], inplace=True)
            
            # Final touch: Add display name for the final excel
            df_export.insert(2, '품목명', df_export['MaterialID'].apply(self.get_material_display_name))
            df_export.to_excel(file_path, index=False)
            messagebox.showinfo("성공", f"파일이 저장되었습니다:\n{file_path}")
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 저장 실패: {e}")

    def update_transaction_view(self):
        for item in self.inout_tree.get_children(): self.inout_tree.delete(item)
        if self.transactions_df.empty: 
            self.inout_summary_var.set("총 0건의 내역")
            return
            
        q = self.inout_search_var.get().strip().lower()
        df = self.transactions_df.sort_values(by='Date', ascending=False)
        
        count = 0
        for _, row in df.iterrows():
            mat_name = self.get_material_display_name(row['MaterialID'])
            site = str(row.get('Site',''))
            t_type = str(row.get('Type',''))
            note = str(row.get('Note',''))
            
            if q and not (q in mat_name.lower() or q in site.lower() or q in t_type.lower() or q in note.lower()):
                continue
            
            vals = (pd.to_datetime(row['Date']).strftime('%Y-%m-%d %H:%M:%S'), site, t_type, mat_name, f"{row.get('Quantity',0):g}", note)
            self.inout_tree.insert('', 'end', values=vals)
            count += 1
            if not q and count >= 100: break # Show only 100 if no search
            
        self.inout_summary_var.set(f"총 {count}건의 내역" + (" (필터링됨)" if q else ""))

    def register_new_material(self, name, warehouse='미지정', 규격='자동등록'):
        mat_id = str(int(self.materials_df['MaterialID'].max()) + 1) if not self.materials_df.empty else "1001"
        new_mat = {'MaterialID': mat_id, '품목명': name, '창고': warehouse, '규격': 규격, '수량': 0, 'Active': 1}
        self.materials_df = pd.concat([self.materials_df, pd.DataFrame([new_mat])], ignore_index=True)
        self.save_data(); self.update_stock_view(); self.update_material_combo()
        return mat_id

    def add_daily_usage_entry(self):
        try:
            date_val = self.ent_daily_date.get_date()
            site = self.cb_daily_site.get().strip()
            if not site: messagebox.showwarning("입력 오류", "현장명을 입력해주세요."); return
            mat_display = self.cb_daily_material.get().strip()
            mat_id = self._get_id_from_display(mat_display) if mat_display else ""
            if not mat_id and mat_display:
                if messagebox.askyesno("신규 자재", f"'{mat_display}'을(는) 자동 등록하시겠습니까?"):
                    mat_id = self.register_new_material(mat_display, warehouse='현장')
            
            saved = self._add_single_usage_record_logic(mat_id, date_val, site)
            if saved > 0:
                self.clear_daily_usage_form_all(keep_date=True)
                self.update_daily_usage_view(); self.update_transaction_view(); self.update_stock_view()
                messagebox.showinfo("성공", "저장되었습니다.")
        except Exception as e: messagebox.showerror("오류", f"저장 실패: {e}")

    def _add_single_usage_record_logic(self, mat_id, date_val, site):
        def to_f(val):
            try: return float(str(val).replace(',','')) if val else 0.0
            except: return 0.0
        
        workers = {}
        for i in range(1, 11):
            g = getattr(self, f'worker_group{i}', None)
            if g:
                workers[f'User{i if i>1 else ""}'] = g.get_worker()
                workers[f'WorkTime{i if i>1 else ""}'] = g.get_time()
                workers[f'OT{i if i>1 else ""}'] = g.get_ot()

        # RTK Defect Data
        rtk_data = {}
        for k, v in self.rtk_vars.items():
            rtk_data[f'RTK_{k}'] = to_f(v.get())

        # NDT Company Logic
        records = []
        common = {
            'Date': date_val, 'Site': site, 'MaterialID': mat_id,
            '검사방법': self.cb_daily_test_method.get(),
            '검사품명': self.ent_daily_inspection_item.get(),
            'Usage': to_f(self.ent_daily_test_amount.get()),
            'Unit': self.cb_daily_unit.get(),
            '업체명': self.cb_daily_company.get(),
            '단가': to_f(self.ent_daily_unit_price.get()),
            '출장비': to_f(self.ent_daily_travel_cost.get()),
            '일식': to_f(self.ent_daily_meal_cost.get()),
            '검사비': to_f(self.ent_daily_test_fee.get()),
            '적용코드': self.ent_daily_applied_code.get(),
            '성적서번호': self.ent_daily_report_no.get(),
            '비고': self.ent_daily_note.get(),
            **workers,
            **rtk_data
        }
        
        for i, entries in enumerate(self.ndt_company_entries):
            row = common.copy()
            row['회사코드'] = entries['_company'].get()
            for k in self.ndt_materials_all:
                row[f'NDT_{k}'] = to_f(entries[k].get())
            if i > 0: # Zero out main fields for secondary companies
                for k in ['Usage','검사량','단가','출장비','일식','검사비']: row[k] = 0
            records.append(row)
            
        self.daily_usage_df = pd.concat([self.daily_usage_df, pd.DataFrame(records)], ignore_index=True)
        self.save_data()
        return len(records)

    def add_ndt_company_section(self):
        idx = len(self.ndt_company_entries)
        frame = ttk.Frame(self.ndt_company_container)
        frame.pack(fill='x', pady=2)
        
        ttk.Label(frame, text=f"업체 {idx+1}:").pack(side='left', padx=2)
        cb_co = ttk.Combobox(frame, width=12, values=self.companies)
        cb_co.pack(side='left', padx=2)
        
        entry_map = {'_frame': frame, '_company': cb_co}
        for mat in self.ndt_materials_all:
            ttk.Label(frame, text=mat[:2]).pack(side='left', padx=1)
            ent = ttk.Entry(frame, width=5)
            ent.pack(side='left', padx=1)
            ent.insert(0, "0")
            entry_map[mat] = ent
            
        self.ndt_company_entries.append(entry_map)

    def remove_last_ndt_company(self):
        if len(self.ndt_company_entries) > 1:
            last = self.ndt_company_entries.pop()
            last['_frame'].destroy()

    def sync_worker_times(self):
        master = self.worker_groups[0]
        m_time = master.ent_worktime.get()
        m_shift = master.cb_shift.get()
        m_ot = master.ent_ot.get()
        for i in range(1, len(self.worker_groups)):
            g = self.worker_groups[i]
            if g.get_worker():
                g.cb_shift.set(m_shift)
                g.ent_worktime.set(m_time)
                g.ent_ot.delete(0, tk.END); g.ent_ot.insert(0, m_ot)

    def clear_daily_usage_form_all(self, keep_date=False):
        for group in self.worker_groups:
            group.set_worker(""); group.set_time(""); group.set_ot("")
        self.cb_daily_material.set(""); self.ent_daily_test_amount.delete(0, tk.END)
        self.ent_daily_unit_price.delete(0, tk.END); self.ent_daily_test_fee.delete(0, tk.END)
        self.ent_daily_travel_cost.delete(0, tk.END); self.ent_daily_travel_cost.insert(0, "0")
        self.ent_daily_meal_cost.delete(0, tk.END); self.ent_daily_meal_cost.insert(0, "0")
        self.ent_daily_note.delete(0, tk.END); self.ent_daily_report_no.delete(0, tk.END)
        # Reset NDT
        while len(self.ndt_company_entries) > 1: self.remove_last_ndt_company()
        if self.ndt_company_entries:
            for k in self.ndt_materials_all: self.ndt_company_entries[0][k].delete(0, tk.END); self.ndt_company_entries[0][k].insert(0, "0")
        # Reset RTK
        for v in self.rtk_vars.values():
            v.delete(0, tk.END); v.insert(0, "0")
        self.update_rtk_total()

    def update_daily_usage_view(self):
        for item in self.daily_tree.get_children(): self.daily_tree.delete(item)
        if self.daily_usage_df.empty: 
            self.daily_summary_var.set("조회된 내역: 0건 | 총 검사비: 0원")
            return
            
        start_date = self.ent_daily_filter_start.get_date()
        end_date = self.ent_daily_filter_end.get_date()
        site_filter = self.cb_daily_filter_site.get()
        user_filter = self.cb_daily_filter_user.get()
        q = self.daily_search_var.get().strip().lower()
        
        df = self.daily_usage_df.copy()
        df['Date'] = pd.to_datetime(df['Date']).dt.date
        
        mask = (df['Date'] >= start_date) & (df['Date'] <= end_date)
        if site_filter != '전체': mask &= (df['Site'] == site_filter)
        
        if user_filter != '전체':
            user_mask = pd.Series(False, index=df.index)
            for i in range(1, 11):
                suffix = str(i) if i > 1 else ""
                col = f'User{suffix}'
                if col in df.columns:
                    user_mask |= (df[col] == user_filter)
            mask &= user_mask
            
        df = df[mask].sort_values(by='Date', ascending=False)
        
        count = 0
        total_fee = 0
        for _, row in df.iterrows():
            date_str = row['Date'].strftime('%Y-%m-%d')
            site = str(row.get('Site', ''))
            item = str(row.get('검사품명', ''))
            qty = row.get('Usage', 0)
            unit = str(row.get('Unit', ''))
            price = row.get('단가', 0)
            try: fee = float(row.get('검사비', 0))
            except: fee = 0
            user = str(row.get('User', ''))
            
            if q and not (q in site.lower() or q in item.lower() or q in user.lower()):
                continue
                
            vals = (date_str, site, item, f"{qty:g}", unit, f"{price:g}", f"{fee:,.0f}", user)
            self.daily_tree.insert('', 'end', values=vals)
            count += 1
            total_fee += fee
            
        self.daily_summary_var.set(f"조회된 내역: {count}건 | 총 검사비: {total_fee:,.0f}원")

    def update_daily_test_fee_calc(self, event=None):
        """Auto-calculate Inspection Fee = (Amount * Unit Price) + Travel Expense (Restored from V13)"""
        try:
            def get_f(entry):
                try:
                    val = entry.get().strip().replace(',', '')
                    return float(val) if val else 0.0
                except: return 0.0

            amount = get_f(self.ent_daily_test_amount)
            price = get_f(self.ent_daily_unit_price)
            travel = get_f(self.ent_daily_travel_cost)
            meal = get_f(self.ent_daily_meal_cost)
            
            calc_fee = (amount * price) + travel + meal
            self.ent_daily_test_fee.delete(0, tk.END)
            self.ent_daily_test_fee.insert(0, f"{calc_fee:,.0f}")
        except: pass

    def save_tab_config(self, force=False):
        """Save current tab configuration to memory and disk (Restored from V13)"""
        try:
            if not getattr(self, 'is_ready', False) and not force: return
            self.root.update_idletasks()
            
            if not hasattr(self, 'tab_config'): self.tab_config = {}
            
            # Simple state capture
            self.tab_config.update({
                'daily_usage_sash_pos': self.daily_usage_paned.sashpos(0) if hasattr(self, 'daily_usage_paned') else None,
                'sites': self.sites,
                'users': self.users,
                'equipments': self.equipments,
                'companies': self.companies,
                'daily_units': self.daily_units
            })
            
            # Save to disk
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(self.tab_config, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            print(f"Failed to save tab config: {e}")

    def load_tab_config(self):
        """Load and restore tab configuration (Restored from V13)"""
        try:
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    self.tab_config = json.load(f)
                
                # Restore sash position
                if 'daily_usage_sash_pos' in self.tab_config and hasattr(self, 'daily_usage_paned'):
                    pos = self.tab_config['daily_usage_sash_pos']
                    if pos: 
                        try: self.daily_usage_paned.sash_place(0, 0, pos) # sashpos can be tricky, using sash_place or wrapping
                        except: pass
                    
        except Exception as e:
            print(f"Failed to load tab config: {e}")

    def open_list_management_dialog(self, title_or_key, data_list=None, config_key=None, target_cb=None):
        """Open a generic dialog to manage list items (Restore from V13)"""
        if data_list is None:
            config_key = title_or_key
            data_map = {
                'sites': ('현장 목록 관리', self.sites),
                'users': ('담당자 목록 관리', self.users),
                'equipments': ('장비 목록 관리', self.equipments),
                'vehicles': ('차량 목록 관리', self.vehicles),
                'companies': ('업체 목록 관리', self.companies),
                'daily_units': ('단위 목록 관리', self.daily_units),
                'test_items': ('검사품명 목록 관리', self.test_methods),
                'applied_codes': ('적용코드 목록 관리', [])
            }
            if config_key not in data_map: return
            title, data_list = data_map[config_key]
        else:
            title = title_or_key
            if config_key is None: config_key = title

        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("500x400")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding=10)
        frame.pack(fill='both', expand=True)
        
        listbox = tk.Listbox(frame, font=('Malgun Gothic', 10))
        listbox.pack(fill='both', expand=True, side='left')
        sb = ttk.Scrollbar(frame, command=listbox.yview); sb.pack(side='right', fill='y')
        listbox.config(yscrollcommand=sb.set)
        
        def refresh_lb():
            listbox.delete(0, tk.END)
            for item in sorted([str(x) for x in data_list if x]):
                listbox.insert(tk.END, item)
        refresh_lb()
        
        btn_f = ttk.Frame(dialog, padding=5)
        btn_f.pack(fill='x')
        
        ent = ttk.Entry(btn_f)
        ent.pack(side='left', fill='x', expand=True, padx=2)
        
        def add_item():
            val = ent.get().strip()
            if val and val not in data_list:
                data_list.append(val)
                ent.delete(0, tk.END); refresh_lb()
                self.refresh_ui_for_list_change(config_key)
        
        def del_item():
            sel = listbox.curselection()
            if sel:
                val = listbox.get(sel[0])
                if val in data_list: data_list.remove(val)
                refresh_lb(); self.refresh_ui_for_list_change(config_key)

        ttk.Button(btn_f, text="추가", command=add_item).pack(side='left', padx=2)
        ttk.Button(btn_f, text="삭제", command=del_item).pack(side='left', padx=2)

    def refresh_ui_for_list_change(self, config_key):
        """Update all UI elements after a list change"""
        # Save to Material_Manager_Config.json if needed
        # For now, let's just update the comboboxes
        if config_key == 'sites':
            vals = sorted(self.sites)
            if hasattr(self, 'cb_daily_site'): self.cb_daily_site['values'] = vals
            if hasattr(self, 'cb_daily_filter_site'): self.cb_daily_filter_site['values'] = ['전체'] + vals
        elif config_key == 'users':
            vals = sorted(self.users)
            for g in self.worker_groups: g.composite.cb_name['values'] = vals
            if hasattr(self, 'cb_daily_filter_user'): self.cb_daily_filter_user['values'] = ['전체'] + vals
        elif config_key == 'companies':
            vals = sorted(self.companies)
            if hasattr(self, 'cb_daily_company'): self.cb_daily_company['values'] = vals
            # Update NDT company combos
            for entry in self.ndt_company_entries: entry['_company']['values'] = vals
        elif config_key == 'daily_units':
            if hasattr(self, 'cb_daily_unit'): self.cb_daily_unit['values'] = sorted(self.daily_units)
        
        # Save settings (assuming save_settings exists or we call save_data)
        self.save_data()
    def open_equipment_search_dialog(self):
        """Search equipment from master (Restored from V13)"""
        dlg = tk.Toplevel(self.root); dlg.title("장비 검색"); dlg.geometry("600x500")
        dlg.transient(self.root); dlg.grab_set()
        
        search_v = tk.StringVar()
        ttk.Entry(dlg, textvariable=search_v).pack(fill='x', padx=10, pady=10)
        
        cols = ('name', 'sn', 'model')
        tree = ttk.Treeview(dlg, columns=cols, show='headings', height=10)
        for c in cols: tree.heading(c, text=c); tree.column(c, width=150)
        tree.pack(fill='both', expand=True, padx=10, pady=5)
        
        def refresh(*args):
            q = search_v.get().strip().lower()
            for i in tree.get_children(): tree.delete(i)
            # Filter for non-consumables (simple heuristic)
            mats = self.materials_df[self.materials_df['품목명'].apply(lambda x: not self._is_consumable_material(str(x), ''))]
            for _, r in mats.iterrows():
                if not q or q in str(r['품목명']).lower() or q in str(r.get('SN','')).lower():
                    tree.insert('', 'end', values=(r['품목명'], r.get('SN',''), r.get('모델명','')))
        
        def select(e=None):
            sel = tree.selection()
            if sel:
                v = tree.item(sel[0])['values']
                self.cb_daily_equip.set(f"{v[0]} ({v[1]})")
                dlg.destroy()

        search_v.trace_add("write", refresh); tree.bind('<Double-1>', select); refresh()

    def open_material_search_dialog(self, target_form='transaction'):
        """Search materials from master (Restored from V13)"""
        dlg = tk.Toplevel(self.root); dlg.title("자재 검색"); dlg.geometry("700x500")
        dlg.transient(self.root); dlg.grab_set()
        
        search_v = tk.StringVar()
        ttk.Entry(dlg, textvariable=search_v).pack(fill='x', padx=10, pady=10)
        
        cols = ('id', 'name', 'model', 'sn', 'stock')
        tree = ttk.Treeview(dlg, columns=cols, show='headings', height=10)
        for c in cols: tree.heading(c, text=c); tree.column(c, width=120)
        tree.pack(fill='both', expand=True, padx=10, pady=5)
        
        def refresh(*args):
            q = search_v.get().strip().lower()
            for i in tree.get_children(): tree.delete(i)
            for _, r in self.materials_df.iterrows():
                if not q or q in str(r['품목명']).lower() or q in str(r.get('모델명','')).lower():
                    tree.insert('', 'end', values=(r['MaterialID'], r['품목명'], r.get('모델명',''), r.get('SN',''), self.calculate_current_stock(r['MaterialID'])))
        
        def select(e=None):
            sel = tree.selection()
            if sel:
                v = tree.item(sel[0])['values']
                disp = self.get_material_display_name(v[0])
                if target_form == 'daily_usage': self.cb_daily_material.set(disp)
                else: self.inout_vars['Material'].set(disp)
                dlg.destroy()

        search_v.trace_add("write", refresh); tree.bind('<Double-1>', select); refresh()

    def _get_id_from_display(self, display):
        for _, row in self.materials_df.iterrows():
            if self.get_material_display_name(row['MaterialID']) == display: return row['MaterialID']
        return None

    def get_material_display_name(self, mat_id):
        m = self.materials_df[self.materials_df['MaterialID'].apply(normalize_id) == normalize_id(mat_id)]
        if m.empty: return str(mat_id)
        row = m.iloc[0]
        return f"[{row.get('창고','?')}] {row.get('품목명','')} ({row.get('SN','')})"

    def save_data(self):
        try:
            with pd.ExcelWriter(self.db_path, engine='openpyxl') as writer:
                self.materials_df.to_excel(writer, sheet_name='Materials', index=False)
                self.transactions_df.to_excel(writer, sheet_name='Transactions', index=False)
                self.daily_usage_df.to_excel(writer, sheet_name='DailyUsage', index=False)
                if self.budget_df is not None: self.budget_df.to_excel(writer, sheet_name='Budget', index=False)
                if self.settings_df is not None: self.settings_df.to_excel(writer, sheet_name='Settings', index=False)
        except Exception as e: messagebox.showerror("저장 오류", f"실패: {e}")

    def select_all_stock(self):
        self.stock_tree.selection_set(self.stock_tree.get_children())

    def show_low_stock(self):
        low = []
        for _, r in self.materials_df.iterrows():
            if r.get('Active', 1) == 0: continue
            curr = self.calculate_current_stock(r['MaterialID'])
            limit = r.get('재고하한', 10)
            if curr < limit: low.append(f"• {r.get('품목명','')}: {curr:g} (하한: {limit:g})")
        if not low: messagebox.showinfo("알림", "재고 부족 항목이 없습니다.")
        else: messagebox.showwarning("재고 부족", "\n".join(low))

    def calculate_current_stock(self, mat_id):
        str_id = normalize_id(mat_id)
        if not str_id: return 0.0
        base_qty = 0.0
        mat_row = self.materials_df[self.materials_df['MaterialID'].apply(normalize_id) == str_id]
        if not mat_row.empty: base_qty = float(mat_row.iloc[0].get('수량', 0))
        trans_qty = 0.0
        if not self.transactions_df.empty:
            m = self.transactions_df['MaterialID'].apply(normalize_id) == str_id
            trans_qty = float(self.transactions_df[m]['Quantity'].sum())
        usage_qty = 0.0
        if not self.daily_usage_df.empty:
            m = self.daily_usage_df['MaterialID'].apply(normalize_id) == str_id
            usage_qty = float(self.daily_usage_df[m]['Usage'].sum())
        return base_qty + trans_qty - usage_qty

    def update_stock_view(self):
        for item in self.stock_tree.get_children(): self.stock_tree.delete(item)
        search = self.search_var.get().lower()
        name_f = self.cb_filter_name.get()
        model_f = self.cb_filter_model.get()
        for _, row in self.materials_df.iterrows():
            if row.get('Active', 1) == 0: continue
            m_name, m_model = str(row.get('품목명','')).lower(), str(row.get('모델명','')).lower()
            if name_f != "전체" and name_f != "" and name_f not in m_name: continue
            if model_f != "전체" and model_f != "" and model_f not in m_model: continue
            if search and (search not in m_name and search not in m_model): continue
            current = self.calculate_current_stock(row['MaterialID'])
            display_vals = (normalize_id(row['MaterialID']), row.get('회사코드',''), row.get('관리품번',''), row.get('품목명',''), row.get('SN',''), row.get('창고',''), row.get('모델명',''), row.get('규격',''), row.get('관리단위',''), f"{current:g}", row.get('재고하한',''))
            self.stock_tree.insert('', 'end', values=display_vals)

    def delete_selected_material(self):
        selected = self.stock_tree.selection()
        if not selected: messagebox.showwarning("선택 오류", "삭제할 품목을 선택해주세요."); return
        if not messagebox.askyesno("삭제 확인", f"선택한 {len(selected)}개 품목을 삭제하시겠습니까?"): return
        for item in selected:
            mat_id = self.stock_tree.item(item, 'values')[0]
            self.materials_df.loc[self.materials_df['MaterialID'].apply(normalize_id) == normalize_id(mat_id), 'Active'] = 0
        self.save_data(); self.update_stock_view(); self.update_material_combo()

    def open_edit_material_dialog(self):
        selected = self.stock_tree.selection()
        if not selected:
            messagebox.showwarning('선택 오류', '편집할 품목을 선택해주세요.')
            return
        
        item_id = self.stock_tree.item(selected[0], 'values')[0]
        mat_row = self.materials_df[self.materials_df['MaterialID'].apply(normalize_id) == normalize_id(item_id)]
        if mat_row.empty:
            messagebox.showerror('오류', '해당 품목을 찾을 수 없습니다.')
            return
            
        row_idx = mat_row.index[0]
        material = mat_row.iloc[0]
        
        dialog = tk.Toplevel(self.root)
        dialog.title('품목 정보 편집')
        dialog.geometry("500x700")
        dialog.grab_set()
        
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(expand=True, fill='both')
        
        # Scrollable area
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)
        
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", expand=True, fill="both")
        scrollbar.pack(side="right", fill="y")
        
        fields = [
            ('MaterialID', 'ID', False),
            ('관리사코드', '관리사코드', True),
            ('관리품목', '관리품목', True),
            ('품목명', '품목명', True),
            ('SN', 'SN', True),
            ('창고', '창고', True),
            ('모델명', '모델명', True),
            ('규격', '규격', True),
            ('품목군코드', '품목군코드', True),
            ('공급업체', '공급업체', True),
            ('제조사', '제조사', True),
            ('제조일', '제조일', True),
            ('가격', '가격', True),
            ('위치', '위치', True),
            ('관리단위', '관리단위', True),
            ('수량', '기초 수량', True),
            ('재고기한', '재고기한', True)
        ]
        
        entries = {}
        for i, (col, label, editable) in enumerate(fields):
            ttk.Label(scroll_frame, text=label).grid(row=i, column=0, sticky='e', padx=5, pady=5)
            ent = ttk.Entry(scroll_frame, width=30)
            val = material.get(col, '')
            if pd.isna(val): val = ''
            ent.insert(0, str(val))
            if not editable: ent.config(state='disabled')
            ent.grid(row=i, column=1, sticky='w', padx=5, pady=5)
            entries[col] = ent
            
        def save():
            for col, ent in entries.items():
                if ent.cget('state') != 'disabled':
                    new_val = ent.get()
                    if col in ['가격', '수량', '재고기한']:
                        try: new_val = float(new_val)
                        except: new_val = 0.0
                    self.materials_df.at[row_idx, col] = new_val
            self.save_data()
            self.update_stock_view()
            dialog.destroy()
            messagebox.showinfo('알림', '수정되었습니다.')
            
        ttk.Button(main_frame, text='저장', command=save).pack(pady=10)
    def _load_ndt_product_map(self):
        try:
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    return json.load(f).get('ndt_product_map', {})
        except: pass
        return {}

    def _save_ndt_product_map(self, map_data):
        try:
            cfg = {}
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f: cfg = json.load(f)
            cfg['ndt_product_map'] = map_data
            with open(self.config_path, 'w', encoding='utf-8') as f: json.dump(cfg, f, indent=4, ensure_ascii=False)
            return True
        except: return False

    def open_ndt_product_map_dialog(self):
        dlg = tk.Toplevel(self.root); dlg.title("NDT 약품-품목 매핑 설정"); dlg.geometry("500x600")
        dlg.transient(self.root); dlg.grab_set()
        main_f = ttk.Frame(dlg, padding=20); main_f.pack(fill='both', expand=True)
        ttk.Label(main_f, text="현장 입력 약품명", font=('Malgun Gothic', 10, 'bold')).grid(row=0, column=0, pady=10)
        ttk.Label(main_f, text="창고 재고 품목 (매핑)", font=('Malgun Gothic', 10, 'bold')).grid(row=0, column=1, pady=10)
        
        current_map = self._load_ndt_product_map()
        options = sorted([self.get_material_display_name(r['MaterialID']) for _, r in self.materials_df.iterrows() if r.get('Active', 1) == 1])
        combos = {}
        for i, mat in enumerate(self.ndt_materials_all):
            ttk.Label(main_f, text=mat).grid(row=i+1, column=0, padx=5, pady=5, sticky='w')
            cb = ttk.Combobox(main_f, values=options, width=40); cb.grid(row=i+1, column=1, padx=5, pady=5, sticky='ew')
            cid = current_map.get(mat, "")
            if cid: cb.set(self.get_material_display_name(cid))
            combos[mat] = cb

        def _save():
            new_map = {}
            for mat, cb in combos.items():
                disp = cb.get().strip()
                if disp:
                    for _, r in self.materials_df.iterrows():
                        if self.get_material_display_name(r['MaterialID']) == disp: new_map[mat] = r['MaterialID']; break
            if self._save_ndt_product_map(new_map): messagebox.showinfo("성공", "저장되었습니다."); dlg.destroy()
            else: messagebox.showerror("오류", "저장 실패")

        btn_f = ttk.Frame(main_f); btn_f.grid(row=len(self.ndt_materials_all)+1, column=0, columnspan=2, pady=20)
        ttk.Button(btn_f, text="저장", command=_save, width=10).pack(side='left', padx=8)
        ttk.Button(btn_f, text="닫기", command=dlg.destroy, width=10).pack(side='left', padx=8)
    def reset_stock_filters(self): self.cb_filter_name.set("전체"); self.cb_filter_model.set("전체"); self.search_var.set(""); self.update_stock_view()
    def export_stock_to_excel(self):
        try:
            path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if not path: return
            data = [self.stock_tree.item(i, 'values') for i in self.stock_tree.get_children()]
            cols = ('ID', '회사코드', '관리품번', '품목명', 'SN', '창고', '모델명', '규격', '단위', '수량', '재고하한')
            pd.DataFrame(data, columns=cols).to_excel(path, index=False)
            messagebox.showinfo("완료", "내보내기 완료")
        except Exception as e: messagebox.showerror("오류", f"실패: {e}")

    def update_material_combo(self):
        mats = [self.get_material_display_name(row['MaterialID']) for _, row in self.materials_df.iterrows() if row.get('Active', 1) == 1]
        if hasattr(self, 'inout_vars') and 'Material' in self.inout_vars:
            self.inout_vars['Material']['values'] = mats
        if hasattr(self, 'cb_daily_material'):
            self.cb_daily_material['values'] = mats
            
        sites = sorted(list(set(self.daily_usage_df['Site'].dropna().tolist() + self.transactions_df['Site'].dropna().tolist())))
        if hasattr(self, 'inout_vars') and 'Site' in self.inout_vars:
            self.inout_vars['Site']['values'] = sites
        if hasattr(self, 'cb_daily_site'):
            self.cb_daily_site['values'] = sites

    def _is_consumable_material(self, n, m): return any(k in n.lower() for k in ['약품','필름','글리세린','페인트','자분'])

    def export_daily_report_to_excel(self):
        try:
            site = self.cb_site_select.get()
            date = self.cal_daily.get_date()
            if not site: messagebox.showwarning("입력 오류", "현장을 선택해주세요."); return
            
            # Find the record
            mask = (self.daily_usage_df['Date'] == date.strftime("%Y-%m-%d")) & (self.daily_usage_df['Site'] == site)
            if self.daily_usage_df[mask].empty:
                messagebox.showwarning("데이터 없음", "해당 날짜와 현장의 저장된 기록이 없습니다."); return
            
            row = self.daily_usage_df[mask].iloc[0]
            
            # Prepare RTK data
            rtk_info = {}
            for k in ['센터미스', '농도', '마킹미스', '필름마크', '취급부주의', '고객불만', '기타']:
                rtk_info[k] = row.get(f'RTK_{k}', 0)
            
            # Prepare OT status
            ot_status = []
            for i in range(1, 11):
                suffix = str(i) if i > 1 else ""
                u = row.get(f'User{suffix}', "")
                t = row.get(f'WorkTime{suffix}', "")
                o = row.get(f'OT{suffix}', "")
                if u:
                    ot_status.append({
                        'names': u,
                        'company': row.get('업체명', '원자력건설'),
                        'method': row.get('검사방법', 'RT'),
                        'ot_hours': t,
                        'ot_amount': o
                    })

            # Prepare data for manager
            data = {
                'date': date,
                'company': row.get('업체명', '원자력건설'),
                'project_name': site,
                'standard': row.get('적용코드', 'KS'),
                'equipment': row.get('장비명', ''),
                'report_no': row.get('성적서번호', ''),
                'inspection_item': row.get('검사품명', ''),
                'inspector': row.get('User', ''),
                'car_no': '',
                'methods': {},
                'rtk': rtk_info,
                'ot_status': ot_status,
                'materials': {},
                'vehicles': [] # Handle vehicle info if needed
            }
            
            # [CRITICAL] Template Path
            template_path = os.path.join(os.path.dirname(__file__), 'resources', 'Template_DailyWorkReport.xlsx')
            if not os.path.exists(template_path):
                # Fallback to absolute if needed for local test
                template_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'resources', 'Template_DailyWorkReport.xlsx')
            
            if not os.path.exists(template_path):
                messagebox.showerror("오류", "엑셀 템플릿 파일을 찾을 수 없습니다."); return
                
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=f"작업일보_{site}_{date.strftime('%Y%m%d')}.xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if not save_path: return
            
            manager = DailyWorkReportManager(template_path)
            manager.generate_report(data, save_path, custom_mapping=self.load_report_mapping())
            
            messagebox.showinfo("완료", "엑셀 보고서가 생성되었습니다.")
            if messagebox.askyesno("열기", "지금 파일을 여시겠습니까?"):
                os.startfile(save_path)
                
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("오류", f"엑셀 출력 중 오류가 발생했습니다: {e}")

    def load_report_mapping(self):
        # Placeholder for real mapping logic from V13
        return None

    def save_report_mapping(self, mapping):
        pass

    def open_report_mapping_dialog(self):
        messagebox.showinfo("알림", "매핑 설정 기능은 현재 개발 중입니다.")

    def import_master_data(self):
        """엑셀 파일에서 자재 데이터 가져오기"""
        file_path = filedialog.askopenfilename(title="엑셀 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path: return
        
        try:
            raw_df = pd.read_excel(file_path)
            # Basic mapping logic (simplified)
            new_mats = []
            for _, row in raw_df.iterrows():
                mat_id = self._generate_material_id()
                new_mats.append({
                    'MaterialID': mat_id,
                    '회사코드': row.get('회사코드', ''),
                    '관리품번': row.get('관리품번', ''),
                    '품목명': row.get('품목명', 'Unknown'),
                    'SN': row.get('SN', ''),
                    '모델명': row.get('모델명', ''),
                    '규격': row.get('규격', ''),
                    '관리단위': row.get('단위', 'EA'),
                    '수량': float(row.get('수량', 0)),
                    '재고하한': float(row.get('재고하한', 0)),
                    'Active': 1
                })
                # Add to materials_df
                self.materials_df = pd.concat([self.materials_df, pd.DataFrame(new_mats)], ignore_index=True)
                new_mats = [] # Reset for next
                
            self.save_data()
            self.update_stock_view()
            self.update_material_combo()
            messagebox.showinfo("완료", "데이터를 성공적으로 가져왔습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"가져오기 실패: {e}")

    def backup_all_data(self):
        """전체 데이터를 새로운 엑셀 파일로 백업"""
        try:
            path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                initialfile=f"Backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                                filetypes=[("Excel files", "*.xlsx")])
            if not path: return
            
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                self.materials_df.to_excel(writer, sheet_name='Materials', index=False)
                self.transactions_df.to_excel(writer, sheet_name='Transactions', index=False)
                self.monthly_usage_df.to_excel(writer, sheet_name='MonthlyUsage', index=False)
                self.daily_usage_df.to_excel(writer, sheet_name='DailyUsage', index=False)
                self.budget_df.to_excel(writer, sheet_name='Budget', index=False)
                if self.settings_df is not None:
                    self.settings_df.to_excel(writer, sheet_name='Settings', index=False)
            
            messagebox.showinfo("완료", f"전체 데이터가 백업되었습니다: {os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("오류", f"백업 실패: {e}")

    # --- UI Helpers for NDT Tabs ---
    def _create_scrollable_sidebar(self, parent):
        canvas = tk.Canvas(parent, background="#f9fafb", highlightthickness=0, borderwidth=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, background="#f9fafb", padx=10, pady=0)
        def _update_scrollregion(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            if canvas.bbox("all")[3] <= canvas.winfo_height(): scrollbar.pack_forget()
            else: scrollbar.pack(side="right", fill="y")
        scrollable_frame.bind("<Configure>", _update_scrollregion)
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        def _on_canvas_configure(e): canvas.itemconfig(canvas_window, width=e.width)
        canvas.bind("<Configure>", _on_canvas_configure)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        return scrollable_frame

    def _create_setting_grid(self, parent, context):
        """Helper to create position/size entries for report objects."""
        items = [
            ("서울로고(Seoul):", f"{context}_SEOUL"), ("SITCO로고(SITCO):", f"{context}_SITCO"),
            ("하부직인(Footer):", f"{context}_FOOTER"), ("하부PT(Footer_PT):", f"{context}_FOOTER_PT")
        ]
        next_row = 0
        for i, (label, key_prefix) in enumerate(items):
            row = i * 2; next_row = row + 2
            ttk.Label(parent, text=label, font=("Malgun Gothic", 9, "bold")).grid(row=row, column=0, sticky='w', pady=(5, 2), columnspan=3)
            # Anchor, W, H, X, Y
            for j, (l, k) in enumerate([("셀:", "ANCHOR"), ("W:", "W"), ("H:", "H"), ("X:", "X"), ("Y:", "Y")]):
                ttk.Label(parent, text=l, font=("Arial", 8)).grid(row=row+1, column=j*2, sticky='e')
                var = tk.StringVar(value=str(self.config.get(f"{key_prefix}_{k}", "")))
                self.setting_vars[f"{key_prefix}_{k}"] = var
                ttk.Entry(parent, textvariable=var, width=6 if k=="ANCHOR" else 5).grid(row=row+1, column=j*2+1, sticky='w', padx=2)
        return next_row

    def _create_margin_settings(self, parent, context, start_row):
        ttk.Label(parent, text="페이지 여백 및 배율:", font=("Malgun Gothic", 9, "bold")).grid(row=start_row, column=0, sticky='w', pady=(10, 2), columnspan=3)
        m_row = start_row + 1
        for i, (l, k) in enumerate([("상:", "TOP"), ("하:", "BOTTOM"), ("좌:", "LEFT"), ("우:", "RIGHT"), ("배율:", "SCALE"), ("영역:", "AREA")]):
            if k == "SCALE": full_key = f"PRINT_SCALE_{context}"
            elif k == "AREA": full_key = f"PRINT_AREA_{context}"
            else: full_key = f"MARGIN_{context}_{k}"
            ttk.Label(parent, text=l, font=("Arial", 8)).grid(row=m_row, column=i*2, sticky='e')
            var = tk.StringVar(value=str(self.config.get(full_key, "")))
            self.setting_vars[full_key] = var
            ttk.Entry(parent, textvariable=var, width=5 if k != "AREA" else 10).grid(row=m_row, column=i*2+1, sticky='w', padx=2)

    def _create_row_settings(self, parent, mode="PMI"):
        """Row range settings for data extraction and printing."""
        ttk.Label(parent, text="데이터 영역 (Row Range):", font=("Malgun Gothic", 9, "bold")).grid(row=0, column=0, sticky='w', pady=(5, 5), columnspan=4)
        items = [("시작 행:", "START_ROW"), ("데이터 종료:", "DATA_END_ROW"), ("출력 종료:", "PRINT_END_ROW")]
        for i, (l, k) in enumerate(items):
            full_key = f"{mode}_{k}"
            ttk.Label(parent, text=l).grid(row=i+1, column=0, sticky='e', pady=2)
            var = tk.StringVar(value=str(self.config.get(full_key, "")))
            self.setting_vars[full_key] = var
            ttk.Entry(parent, textvariable=var, width=10).grid(row=i+1, column=1, sticky='w', padx=5)

    def _create_column_mapping_ui(self, parent, mode, items):
        """Standard column mapping grid."""
        canvas = tk.Canvas(parent, background="#f9fafb", highlightthickness=0); canvas.pack(side="left", fill="both", expand=True)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview); vsb.pack(side="right", fill="y")
        f = tk.Frame(canvas, background="#f9fafb"); canvas.create_window((0,0), window=f, anchor='nw'); canvas.configure(yscrollcommand=vsb.set)
        def _on_f_cfg(e): canvas.configure(scrollregion=canvas.bbox("all"))
        f.bind("<Configure>", _on_f_cfg)
        ttk.Label(f, text="항목", font=("Malgun Gothic", 9, "bold")).grid(row=0, column=0, sticky='w')
        ttk.Label(f, text="Excel 열(숫자)", font=("Malgun Gothic", 9, "bold")).grid(row=0, column=1, sticky='w', padx=10)
        ttk.Label(f, text="성적서 표기명", font=("Malgun Gothic", 9, "bold")).grid(row=0, column=2, sticky='w')
        for i, (label, col_key, def_idx, name_key, def_name, placeholder) in enumerate(items):
            row = i + 1
            ttk.Label(f, text=label).grid(row=row, column=0, sticky='e', pady=2)
            var_idx = tk.StringVar(value=str(self.config.get(col_key, def_idx)))
            self.setting_vars[col_key] = var_idx
            ttk.Entry(f, textvariable=var_idx, width=8, justify='center').grid(row=row, column=1, padx=10)
            var_name = tk.StringVar(value=str(self.config.get(name_key, def_name)))
            self.setting_vars[name_key] = var_name
            ttk.Entry(f, textvariable=var_name, width=20).grid(row=row, column=2)

    def _create_preview_ui(self, parent, mode="PMI"):
        """Creates the main live preview treeview with search/filter/actions."""
        container = tk.Frame(parent, background="#f9fafb"); container.pack(fill='both', expand=True)
        # Search & Sort Filter Area
        ctrl = tk.Frame(container, background="#f9fafb"); ctrl.pack(fill='x', pady=(0, 5))
        ttk.Label(ctrl, text="🔍 검색:").pack(side='left', padx=2)
        ent_search = ttk.Entry(ctrl, width=20); ent_search.pack(side='left', padx=2)
        ttk.Button(ctrl, text="필터 적용", command=lambda: self.apply_preview_filter(mode)).pack(side='left', padx=5)
        
        # Treeview
        tree_frame = tk.Frame(container); tree_frame.pack(fill='both', expand=True)
        cols = ["V", "No", "Date", "Dwg No", "Joint No", "Loc", "Ni", "Cr", "Mo", "Grade"] if mode=="PMI" else ["V", "No", "Date", "ISO/Dwg", "Joint No.", "Result", "Welder", "Remarks"]
        tree = ttk.Treeview(tree_frame, columns=cols, show='headings', selectmode='extended')
        for c in cols: tree.heading(c, text=c); tree.column(c, width=80, anchor='center')
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview); tree.configure(yscrollcommand=vsb.set)
        tree.pack(side='left', fill='both', expand=True); vsb.pack(side='right', fill='y')
        if mode=="PMI": self.preview_tree = tree
        elif mode=="RT": self.rt_preview_tree = tree
        elif mode=="PT": self.pt_preview_tree = tree
        elif mode=="PAUT": self.paut_preview_tree = tree

    def apply_preview_filter(self, mode):
        """Filters the current preview treeview based on search criteria."""
        messagebox.showinfo("알림", f"{mode} 필터 기능이 준비 중입니다.")

    def setup_pmi_tab(self, parent):
        container = tk.Frame(parent, background="#f9fafb"); container.pack(fill='both', expand=True)
        pw = tk.PanedWindow(container, orient='horizontal', background="#d1d5db", sashwidth=4); pw.pack(fill='both', expand=True)
        left = tk.Frame(pw, background="#f9fafb"); pw.add(left, width=400)
        lp = self._create_scrollable_sidebar(left)
        ttk.Label(lp, text="🔬 PMI 데이터 관리", font=("Malgun Gothic", 12, "bold")).pack(anchor='w', pady=10)
        
        # File Selection
        f_box = ttk.LabelFrame(lp, text=" 파일 선택 "); f_box.pack(fill='x', pady=5)
        self._add_compact_file_row(f_box, "데이터:", self.target_file_path, 0)
        self._add_compact_file_row(f_box, "양식:", self.template_file_path, 1)
        
        # Config Notebook
        nb = ttk.Notebook(lp); nb.pack(fill='both', expand=True, pady=10)
        t_cover = ttk.Frame(nb, padding=5); nb.add(t_cover, text="갑지")
        t_data = ttk.Frame(nb, padding=5); nb.add(t_data, text="을지")
        t_rows = ttk.Frame(nb, padding=5); nb.add(t_rows, text="행")
        t_cols = ttk.Frame(nb, padding=5); nb.add(t_cols, text="열")
        self._create_setting_grid(t_cover, "PMI_COVER")
        self._create_setting_grid(t_data, "PMI_DATA")
        self._create_row_settings(t_rows, "PMI")
        
        ttk.Button(lp, text=" ✨ 생성 시작 ", style="Action.TButton", command=self.run_process).pack(fill='x', pady=5)
        
        right = ttk.LabelFrame(pw, text=" 실시간 데이터 미리보기 "); pw.add(right, stretch="always")
        self._create_preview_ui(right, "PMI")

    def _add_compact_file_row(self, parent, label, var, row):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky='e', padx=2)
        ttk.Entry(parent, textvariable=var, width=1).grid(row=row, column=1, sticky='ew', padx=2)
        ttk.Button(parent, text="...", width=3, command=lambda: self._browse_file(var)).grid(row=row, column=2, padx=2)
        parent.columnconfigure(1, weight=1)

    def _browse_file(self, var):
        path = filedialog.askopenfilename()
        if path: var.set(path)


    def setup_rt_tab(self, parent):
        container = tk.Frame(parent, background="#f9fafb"); container.pack(fill='both', expand=True)
        pw = tk.PanedWindow(container, orient='horizontal', background="#d1d5db", sashwidth=4); pw.pack(fill='both', expand=True)
        left = tk.Frame(pw, background="#f9fafb"); pw.add(left, width=400)
        lp = self._create_scrollable_sidebar(left)
        ttk.Label(lp, text="🔬 RT 데이터 관리", font=("Malgun Gothic", 12, "bold")).pack(anchor='w', pady=10)
        f_box = ttk.LabelFrame(lp, text=" 파일 선택 "); f_box.pack(fill='x', pady=5)
        self._add_compact_file_row(f_box, "데이터:", self.rt_target_file_path, 0)
        self._add_compact_file_row(f_box, "양식:", self.rt_template_file_path, 1)
        ttk.Button(f_box, text="💾 현재 설정을 이 양식 전용으로 저장", command=lambda: self.save_template_specific_config("RT")).grid(row=2, column=1, sticky='w', padx=5, pady=2)
        nb = ttk.Notebook(lp); nb.pack(fill='both', expand=True, pady=10)
        t_cover = ttk.Frame(nb, padding=5); nb.add(t_cover, text="갑지")
        t_data = ttk.Frame(nb, padding=5); nb.add(t_data, text="을지")
        t_rows = ttk.Frame(nb, padding=5); nb.add(t_rows, text="행")
        t_cols = ttk.Frame(nb, padding=5); nb.add(t_cols, text="열")
        self._create_setting_grid(t_cover, "RT_COVER")
        self._create_setting_grid(t_data, "RT_DATA")
        self._create_margin_settings(t_cover, "RT_COVER", 5) # Added Margin/Area/Scale
        self._create_margin_settings(t_data, "RT_DATA", 5)
        self._create_row_settings(t_rows, "RT")
        rt_items = [
            ("No:", "RT_COL_NO", 1, "RT_NAME_NO", "No", "No"),
            ("Date:", "RT_COL_DATE", 2, "RT_NAME_DATE", "Date", "Date"),
            ("Drawing No:", "RT_COL_DWG", 3, "RT_NAME_DWG", "Drawing No.", "Drawing No."),
            ("Joint No:", "RT_COL_JOINT", 4, "RT_NAME_JOINT", "Joint No.", "Joint No."),
            ("Location:", "RT_COL_LOC", 5, "RT_NAME_LOC", "Location", "Location"),
            ("Result:", "RT_COL_RES", 28, "RT_NAME_RES", "Result", "Result")
        ]
        self._create_column_mapping_ui(t_cols, "RT", rt_items)
        ttk.Button(lp, text=" ✨ 생성 시작 ", style="Action.TButton", command=self.run_process).pack(fill='x', pady=5)
        right = ttk.LabelFrame(pw, text=" RT 데이터 미리보기 "); pw.add(right, stretch="always")
        self._create_preview_ui(right, "RT")

    def setup_pt_tab(self, parent):
        container = tk.Frame(parent, background="#f9fafb"); container.pack(fill='both', expand=True)
        pw = tk.PanedWindow(container, orient='horizontal', background="#d1d5db", sashwidth=4); pw.pack(fill='both', expand=True)
        left = tk.Frame(pw, background="#f9fafb"); pw.add(left, width=400)
        lp = self._create_scrollable_sidebar(left)
        ttk.Label(lp, text="🔬 PT 데이터 관리", font=("Malgun Gothic", 12, "bold")).pack(anchor='w', pady=10)
        f_box = ttk.LabelFrame(lp, text=" 파일 선택 "); f_box.pack(fill='x', pady=5)
        self._add_compact_file_row(f_box, "데이터:", self.pt_target_file_path, 0)
        self._add_compact_file_row(f_box, "양식:", self.pt_template_file_path, 1)
        ttk.Button(f_box, text="💾 현재 설정을 이 양식 전용으로 저장", command=lambda: self.save_template_specific_config("PT")).grid(row=2, column=1, sticky='w', padx=5, pady=2)
        nb = ttk.Notebook(lp); nb.pack(fill='both', expand=True, pady=10)
        t_cover = ttk.Frame(nb, padding=5); nb.add(t_cover, text="갑지")
        t_data = ttk.Frame(nb, padding=5); nb.add(t_data, text="을지")
        t_rows = ttk.Frame(nb, padding=5); nb.add(t_rows, text="행")
        t_cols = ttk.Frame(nb, padding=5); nb.add(t_cols, text="열")
        self._create_setting_grid(t_cover, "PT_COVER")
        self._create_setting_grid(t_data, "PT_DATA")
        self._create_margin_settings(t_cover, "PT_COVER", 5) # Added Margin/Area/Scale
        self._create_margin_settings(t_data, "PT_DATA", 5)
        self._create_row_settings(t_rows, "PT")
        pt_items = [
            ("No:", "PT_COL_NO", 1, "PT_NAME_NO", "No", "No"),
            ("Dwg:", "PT_COL_DWG", 2, "PT_NAME_DWG", "Dwg", "Dwg"),
            ("Joint:", "PT_COL_JOINT", 5, "PT_NAME_JOINT", "Joint", "Joint"),
            ("Result:", "PT_COL_RES", 11, "PT_NAME_RES", "Result", "Result")
        ]
        self._create_column_mapping_ui(t_cols, "PT", pt_items)
        ttk.Button(lp, text=" ✨ 생성 시작 ", style="Action.TButton", command=self.run_process).pack(fill='x', pady=5)
        right = ttk.LabelFrame(pw, text=" PT 데이터 미리보기 "); pw.add(right, stretch="always")
        self._create_preview_ui(right, "PT")

    def setup_paut_tab(self, parent):
        container = tk.Frame(parent, background="#f9fafb"); container.pack(fill='both', expand=True)
        pw = tk.PanedWindow(container, orient='horizontal', background="#d1d5db", sashwidth=4); pw.pack(fill='both', expand=True)
        left = tk.Frame(pw, background="#f9fafb"); pw.add(left, width=400)
        lp = self._create_scrollable_sidebar(left)
        ttk.Label(lp, text="🔬 PAUT 데이터 관리", font=("Malgun Gothic", 12, "bold")).pack(anchor='w', pady=10)
        f_box = ttk.LabelFrame(lp, text=" 파일 선택 "); f_box.pack(fill='x', pady=5)
        self._add_compact_file_row(f_box, "데이터:", self.paut_target_file_path, 0)
        self._add_compact_file_row(f_box, "양식:", self.paut_template_file_path, 1)
        nb = ttk.Notebook(lp); nb.pack(fill='both', expand=True, pady=10)
        t_cover = ttk.Frame(nb, padding=5); nb.add(t_cover, text="갑지")
        t_data = ttk.Frame(nb, padding=5); nb.add(t_data, text="을지")
        t_rows = ttk.Frame(nb, padding=5); nb.add(t_rows, text="행")
        t_cols = ttk.Frame(nb, padding=5); nb.add(t_cols, text="열")
        self._create_setting_grid(t_cover, "PAUT_COVER")
        self._create_setting_grid(t_data, "PAUT_DATA")
        self._create_row_settings(t_rows, "PAUT")
        paut_items = [
            ("No:", "PAUT_COL_NO", 1, "PAUT_NAME_NO", "No", "No"),
            ("Line No:", "PAUT_COL_LINE", 2, "PAUT_NAME_LINE", "Line No.", "Line No."),
            ("Joint No:", "PAUT_COL_JOINT", 3, "PAUT_NAME_JOINT", "Joint No.", "Joint No."),
            ("Evaluation:", "PAUT_COL_EVAL", 9, "PAUT_NAME_EVAL", "Evaluation", "Evaluation")
        ]
        self._create_column_mapping_ui(t_cols, "PAUT", paut_items)
        ttk.Button(lp, text=" ✨ 생성 시작 ", style="Action.TButton", command=self.run_process).pack(fill='x', pady=5)
        right = ttk.LabelFrame(pw, text=" PAUT 데이터 미리보기 "); pw.add(right, stretch="always")
        self._create_preview_ui(right, "PAUT")

    def setup_photo_log_tab(self, parent):
        """Standardized Photo Log UI with Dual-Pane Layout."""
        self.photo_paned = tk.PanedWindow(parent, orient='horizontal', background="#d1d5db", 
                            sashwidth=6, sashpad=0, sashrelief='raised', borderwidth=0)
        self.photo_paned.pack(fill='both', expand=True)

        # [LEFT] Settings Sidebar
        left_container = tk.Frame(self.photo_paned, background="#f9fafb", highlightthickness=0, borderwidth=0)
        self.photo_paned.add(left_container, width=425, minsize=200)
        
        # Scrollable area
        left_pane = self._create_scrollable_sidebar(left_container)

        # 1. Report Info Group
        info_frame = ttk.LabelFrame(left_pane, text=" 리포트 정보 (Report Info) ", padding=10)
        info_frame.pack(fill='x', padx=10, pady=5)

        tk.Label(info_frame, text="검사 항목:").grid(row=0, column=0, sticky='w', pady=2)
        type_combo = ttk.Combobox(info_frame, textvariable=self.photo_inspect_type, 
                                 values=list(self.photo_header_map.keys()), state="readonly")
        type_combo.grid(row=0, column=1, sticky='ew', padx=5, pady=2)
        type_combo.bind("<<ComboboxSelected>>", self._on_photo_type_change)

        tk.Label(info_frame, text="리포트 제목:").grid(row=1, column=0, sticky='w', pady=2)
        ttk.Entry(info_frame, textvariable=self.photo_report_title).grid(row=1, column=1, sticky='ew', padx=5, pady=2)

        tk.Label(info_frame, text="발주처:").grid(row=2, column=0, sticky='w', pady=2)
        ttk.Entry(info_frame, textvariable=self.photo_orderer).grid(row=2, column=1, sticky='ew', padx=5, pady=2)

        tk.Label(info_frame, text="리포트 번호:").grid(row=3, column=0, sticky='w', pady=2)
        ttk.Entry(info_frame, textvariable=self.photo_report_no).grid(row=3, column=1, sticky='ew', padx=5, pady=2)

        tk.Label(info_frame, text="검사 일자:").grid(row=4, column=0, sticky='w', pady=2)
        ttk.Entry(info_frame, textvariable=self.photo_inspect_date).grid(row=4, column=1, sticky='ew', padx=5, pady=2)
        
        tk.Label(info_frame, text="로고 파일:").grid(row=5, column=0, sticky='w', pady=2)
        logo_f = tk.Frame(info_frame)
        logo_f.grid(row=5, column=1, sticky='ew', padx=5, pady=2)
        ttk.Entry(logo_f, textvariable=self.photo_logo_path).pack(side='left', fill='x', expand=True)
        def browse_logo():
            f = filedialog.askopenfilename(filetypes=[("Images", "*.png;*.jpg;*.jpeg;*.bmp")])
            if f: self.photo_logo_path.set(f)
        ttk.Button(logo_f, text="...", width=3, command=browse_logo).pack(side='right')
        
        info_frame.columnconfigure(1, weight=1)

        # 2. Layout Settings Group
        layout_frame = ttk.LabelFrame(left_pane, text=" 사진 레이아웃 설정 (Layout) ", padding=10)
        layout_frame.pack(fill='x', padx=10, pady=5)

        tk.Label(layout_frame, text="한 줄당 사진:").grid(row=0, column=0, sticky='w', pady=2)
        ttk.Combobox(layout_frame, textvariable=self.photo_cols_per_row, values=["1", "2", "3"], state="readonly", width=5).grid(row=0, column=1, sticky='w', padx=5, pady=2)
        ttk.Checkbutton(layout_frame, text="비율 유지", variable=self.photo_keep_aspect).grid(row=0, column=2, sticky='w')

        tk.Label(layout_frame, text="칸 너비/높이:").grid(row=1, column=0, sticky='w', pady=2)
        wh_f = tk.Frame(layout_frame)
        wh_f.grid(row=1, column=1, columnspan=2, sticky='w')
        ttk.Entry(wh_f, textvariable=self.photo_cell_width_var, width=6).pack(side='left', padx=2)
        ttk.Entry(wh_f, textvariable=self.photo_cell_height_var, width=6).pack(side='left', padx=2)

        tk.Label(layout_frame, text="여백(T/B/L/R):").grid(row=2, column=0, sticky='w', pady=2)
        m_f = tk.Frame(layout_frame)
        m_f.grid(row=2, column=1, columnspan=2, sticky='w')
        ttk.Entry(m_f, textvariable=self.photo_margin_top_var, width=4).pack(side='left', padx=1)
        ttk.Entry(m_f, textvariable=self.photo_margin_bottom_var, width=4).pack(side='left', padx=1)
        ttk.Entry(m_f, textvariable=self.photo_margin_left_var, width=4).pack(side='left', padx=1)
        ttk.Entry(m_f, textvariable=self.photo_margin_right_var, width=4).pack(side='left', padx=1)

        tk.Label(layout_frame, text="설명 높이:").grid(row=3, column=0, sticky='w', pady=2)
        ttk.Entry(layout_frame, textvariable=self.photo_desc_height_var, width=10).grid(row=3, column=1, sticky='w', padx=5, pady=2)
        
        tk.Label(layout_frame, text="인쇄 배율:").grid(row=4, column=0, sticky='w', pady=2)
        ttk.Entry(layout_frame, textvariable=self.photo_print_scale_var, width=10).grid(row=4, column=1, sticky='w', padx=5, pady=2)

        tk.Label(layout_frame, text="배치 설정:").grid(row=5, column=0, sticky='w', pady=2)
        b_f = tk.Frame(layout_frame)
        b_f.grid(row=5, column=1, columnspan=2, sticky='w')
        ttk.Combobox(b_f, textvariable=self.photo_align_var, values=["좌측 정렬", "중앙 정렬"], state="readonly", width=10).pack(side='left', padx=2)
        ttk.Checkbutton(b_f, text="가로 폭 맞춤 (Fit to Width)", variable=self.photo_fit_width_var).pack(side='left', padx=5)
        ttk.Checkbutton(b_f, text="세로 사진 자동 회전", variable=self.photo_auto_rotate_var).pack(side='left', padx=5)
        
        tk.Label(layout_frame, text="너비비율(%):").grid(row=6, column=0, sticky='w')
        wf_f = tk.Frame(layout_frame)
        wf_f.grid(row=6, column=1, columnspan=2, sticky='w')
        ttk.Entry(wf_f, textvariable=self.photo_width_pct_var, width=7).pack(side='left', padx=2)
        tk.Label(wf_f, text="너비추가(px):").pack(side='left', padx=(10, 0))
        ttk.Entry(wf_f, textvariable=self.photo_width_pixel_adj_var, width=4).pack(side='left', padx=2)
        tk.Label(wf_f, text="좌우(px):").pack(side='left', padx=(10, 0))
        ttk.Entry(wf_f, textvariable=self.photo_shift_x_var, width=4).pack(side='left', padx=2)
        tk.Label(wf_f, text="상하(px):").pack(side='left', padx=(10, 0))
        ttk.Entry(wf_f, textvariable=self.photo_shift_y_var, width=4).pack(side='left', padx=2)

        # 3. Logo Options
        logo_frame = ttk.LabelFrame(left_pane, text=" 로고 및 출력 설정 ", padding=10)
        logo_frame.pack(fill='x', padx=10, pady=5)

        tk.Label(logo_frame, text="로고 너비:").grid(row=0, column=0, sticky='w', pady=2)
        ttk.Entry(logo_frame, textvariable=self.photo_logo_width_var, width=10).grid(row=0, column=1, sticky='w', padx=5, pady=2)
        
        tk.Label(logo_frame, text="로고 X/Y:").grid(row=1, column=0, sticky='w', pady=2)
        xy_f = tk.Frame(logo_frame)
        xy_f.grid(row=1, column=1, sticky='w')
        ttk.Entry(xy_f, textvariable=self.photo_logo_x_var, width=5).pack(side='left', padx=2)
        ttk.Entry(xy_f, textvariable=self.photo_logo_y_var, width=5).pack(side='left', padx=2)
        
        # [NEW] Photo Log Layout Save Button
        btn_f = tk.Frame(logo_frame, background="#ffffff")
        btn_f.grid(row=2, column=0, columnspan=2, sticky='ew', pady=5)
        ttk.Button(btn_f, text="💾 사진대장 레이아웃 설정 저장", command=self.save_photo_log_config).pack(side='left', padx=5)
        tk.Label(btn_f, text="* 여백, 로고위치 등이 저장됩니다.", font=("Malgun Gothic", 8), fg="gray", background="#ffffff").pack(side='left')
        
        tk.Label(logo_frame, text="출력 파일명:").grid(row=2, column=0, sticky='w', pady=2)
        ttk.Entry(logo_frame, textvariable=self.photo_output_name).grid(row=2, column=1, sticky='ew', padx=5, pady=2)
        logo_frame.columnconfigure(1, weight=1)

        ttk.Button(left_pane, text="🚀 사진대장 리포트 생성", style="Accent.TButton", command=self.start_photo_generation).pack(fill='x', padx=20, pady=10)
        ttk.Button(left_pane, text="💾 현재 설정 저장", command=self.manual_save_settings).pack(fill='x', padx=20, pady=5)

        # [RIGHT] Preview & File List
        right_container = tk.Frame(self.photo_paned, background="#ffffff")
        self.photo_paned.add(right_container)
        
        # File List Header
        list_header = tk.Frame(right_container, background="#f8fafc", padx=15, pady=10)
        list_header.pack(fill='x')
        tk.Label(list_header, text="📁 선택된 사진 리스트 (파일 순차 정렬됨)", font=("Malgun Gothic", 10, "bold"), 
                 background="#f8fafc", foreground="#475569").pack(side='left')
        
        tool_bar = tk.Frame(right_container, background="#ffffff", padx=10, pady=5); tool_bar.pack(fill='x')
        ttk.Button(tool_bar, text="파일 개별 추가", command=self._add_photo_files).pack(side='left', padx=2)
        ttk.Button(tool_bar, text="폴더 전체 추가", command=self._add_photo_folder).pack(side='left', padx=2)
        ttk.Button(tool_bar, text="전체 비우기", command=self._clear_photo_all).pack(side='right', padx=2)
        ttk.Button(tool_bar, text="선택 항목 제거", command=self._remove_photo_selected).pack(side='right', padx=2)

        self.photo_listbox = tk.Listbox(right_container, font=("Consolas", 9), selectmode="extended")
        self.photo_listbox.pack(fill='both', expand=True, padx=10, pady=5)

    def start_photo_generation(self):
        if not self.photo_selected_files:
            messagebox.showwarning("경고", "사진을 선택해주세요.")
            return
        threading.Thread(target=self.generate_photo_report, daemon=True).start()

    def _on_photo_type_change(self, event=None):
        new_type = self.photo_inspect_type.get()
        if new_type in self.photo_header_map:
            self.photo_report_title.set(self.photo_header_map[new_type])

    def _add_photo_files(self):
        files = filedialog.askopenfilenames(filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp")])
        if files:
            for f in files:
                f_norm = os.path.normpath(f)
                if f_norm not in self.photo_selected_files:
                    self.photo_selected_files.append(f_norm)
                    self.photo_listbox.insert(tk.END, f_norm)

    def _add_photo_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            for ext in ('*.png', '*.jpg', '*.jpeg'):
                for f in glob.glob(os.path.join(folder, ext)):
                    f_norm = os.path.normpath(f)
                    if f_norm not in self.photo_selected_files:
                        self.photo_selected_files.append(f_norm)
                        self.photo_listbox.insert(tk.END, f_norm)

    def _remove_photo_selected(self):
        idxs = list(self.photo_listbox.curselection())
        for i in reversed(idxs):
            path = self.photo_listbox.get(i)
            if path in self.photo_selected_files: self.photo_selected_files.remove(path)
            self.photo_listbox.delete(i)

    def manual_save_settings(self):
        """Update config from UI vars and save to file."""
        self.config['PHOTO_ORDERER'] = self.photo_orderer.get()
        self.config['PHOTO_REPORT_NO'] = self.photo_report_no.get()
        self.config['PHOTO_REPORT_TITLE'] = self.photo_report_title.get()
        self.config['PHOTO_INSPECT_TYPE'] = self.photo_inspect_type.get()
        self.config['PHOTO_LOGO_PATH'] = self.photo_logo_path.get()
        self.config['PHOTO_COLS_PER_ROW'] = self.photo_cols_per_row.get()
        self.config['PHOTO_KEEP_ASPECT'] = self.photo_keep_aspect.get()
        self.config['PHOTO_CELL_WIDTH'] = self.photo_cell_width_var.get()
        self.config['PHOTO_CELL_HEIGHT'] = self.photo_cell_height_var.get()
        self.config['PHOTO_MARGIN_TOP'] = self.photo_margin_top_var.get()
        self.config['PHOTO_MARGIN_BOTTOM'] = self.photo_margin_bottom_var.get()
        self.config['PHOTO_MARGIN_LEFT'] = self.photo_margin_left_var.get()
        self.config['PHOTO_MARGIN_RIGHT'] = self.photo_margin_right_var.get()
        self.config['PHOTO_DESC_HEIGHT'] = self.photo_desc_height_var.get()
        self.config['PHOTO_PRINT_SCALE'] = self.photo_print_scale_var.get()
        self.config['PHOTO_ALIGN'] = self.photo_align_var.get()
        self.config['PHOTO_FIT_WIDTH'] = self.photo_fit_width_var.get()
        self.config['PHOTO_AUTO_ROTATE'] = self.photo_auto_rotate_var.get()
        self.config['PHOTO_WIDTH_PCT'] = self.photo_width_pct_var.get()
        self.config['PHOTO_WIDTH_ADJ'] = self.photo_width_pixel_adj_var.get()
        self.config['PHOTO_SHIFT_X'] = self.photo_shift_x_var.get()
        self.config['PHOTO_SHIFT_Y'] = self.photo_shift_y_var.get()
        self.config['PHOTO_LOGO_WIDTH'] = self.photo_logo_width_var.get()
        self.config['PHOTO_LOGO_X'] = self.photo_logo_x_var.get()
        self.config['PHOTO_LOGO_Y'] = self.photo_logo_y_var.get()
        self.config['PHOTO_OUTPUT_NAME'] = self.photo_output_name.get()
        self.config['LAST_PHOTO_SAVE_DIR'] = self.last_photo_save_dir
        
        # [NEW] Capture RT/PT/PMI setting_vars
        if hasattr(self, 'setting_vars'):
            for key, var in self.setting_vars.items():
                val = var.get()
                try:
                    if "AREA" in key:
                        self.config[key] = str(val).strip()
                    elif any(x in key for x in ['_X', '_Y', '_W', '_H', 'MARGIN', 'HEIGHT', 'WIDTH']):
                        self.config[key] = float(val)
                    elif 'SCALE' in key or key.endswith('_ROW'):
                        self.config[key] = int(float(val))
                    else:
                        self.config[key] = str(val)
                except:
                    self.config[key] = str(val)

        self.save_settings()

    def _clear_photo_all(self):
        self.photo_selected_files.clear()
        self.photo_listbox.delete(0, tk.END)

    def generate_photo_report(self):
        try:
            if not self.photo_selected_files: return
            
            image_files = sorted(self.photo_selected_files)
            initial_folder = self.last_photo_save_dir if self.last_photo_save_dir and os.path.exists(self.last_photo_save_dir) else os.path.dirname(image_files[0])
            default_name = self.photo_output_name.get()
            if not default_name.endswith(".xlsx"): default_name += ".xlsx"
            
            # [NEW] Ask for save folder and filename
            output_path = filedialog.asksaveasfilename(
                title="사진대장 저장 위치 선택",
                initialdir=initial_folder,
                initialfile=default_name,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not output_path:
                self.status_var.set("사진대장 생성 취소됨")
                return

            # [REMEMBER] Update last save directory
            self.last_photo_save_dir = os.path.dirname(output_path)
            self.config['LAST_PHOTO_SAVE_DIR'] = self.last_photo_save_dir
            self.save_settings()

            self.progress["value"] = 0
            self.log("[PhotoLog] 작업을 시작합니다..")
            self.status_var.set("사진대장 생성 중...")
            
            workbook = xlsxwriter.Workbook(output_path)
            worksheet = workbook.add_worksheet()

            # Page Setup
            worksheet.set_paper(9) # A4
            worksheet.set_portrait()
            worksheet.center_horizontally()
            
            try:
                m_t = float(self.photo_margin_top_var.get())
                m_b = float(self.photo_margin_bottom_var.get())
                m_l = float(self.photo_margin_left_var.get())
                m_r = float(self.photo_margin_right_var.get())
                worksheet.set_margins(left=m_l, right=m_r, top=m_t, bottom=m_b)
            except:
                worksheet.set_margins(left=0.4, right=0.4, top=0.5, bottom=0.5)
            
            worksheet.set_footer('&C&P / &N')
            worksheet.repeat_rows(0, 4) 

            # Layout Calculation
            num_cols = int(self.photo_cols_per_row.get())
            photos_per_page = 4 if num_cols == 1 else (8 if num_cols == 2 else 12)
            total_pages = math.ceil(len(self.photo_selected_files) / photos_per_page)
            worksheet.fit_to_pages(1, total_pages)

            # Formats
            title_format = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'shrink': True})
            company_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': 9, 'text_wrap': True})
            center_border = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': 10})
            bold_format = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
            desc_format = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'border': 1, 'font_size': 10, 'shrink': True, 'text_wrap': False, 'indent': 1})

            # Fixed 6-column Grid System
            GRID_COLS = 6
            unit_per_grid = (float(self.photo_cell_width_var.get()) * 2) / GRID_COLS
            worksheet.set_column(0, GRID_COLS - 1, unit_per_grid)

            CELL_ROW_HEIGHT = float(self.photo_cell_height_var.get())
            
            try:
                WIDTH_PCT = float(self.photo_width_pct_var.get().strip()) / 100.0
                PIXEL_ADJ = float(self.photo_width_pixel_adj_var.get().strip())
                SHIFT_X = float(self.photo_shift_x_var.get().strip())
                SHIFT_Y = float(self.photo_shift_y_var.get().strip())
            except:
                WIDTH_PCT, PIXEL_ADJ, SHIFT_X, SHIFT_Y = 1.0, 0.0, 0.0, 0.0
                
            INTERNAL_FACTOR = 7.142
            if num_cols == 1:
                photo_col_spans = [(0, GRID_COLS - 1)]
                CELL_WIDTH_PX = round(((unit_per_grid * INTERNAL_FACTOR + 5) * 6), 1)
            elif num_cols == 2:
                photo_col_spans = [(0, 2), (3, 5)]
                CELL_WIDTH_PX = round(((unit_per_grid * INTERNAL_FACTOR + 5) * 3), 1)
            else: # 3 Columns
                photo_col_spans = [(0, 1), (2, 3), (4, 5)]
                CELL_WIDTH_PX = round(((unit_per_grid * INTERNAL_FACTOR + 5) * 2), 1)
            
            SAFE_WIDTH = (CELL_WIDTH_PX - 10) * WIDTH_PCT + PIXEL_ADJ
            
            worksheet.set_row(0, 30)
            worksheet.merge_range(0, 0, 0, GRID_COLS-1, self.photo_report_title.get(), title_format)
            
            company_text = "서   울   檢   査   株   式   會   社\nSEOUL INSPECTION & TESTING Co., Ltd.\nTEL : (02) 552-1112   FAX : (02) 2058-0720"
            worksheet.merge_range(1, 0, 3, 2, company_text, company_format)

            # Logo Insertion: Priority to User Selection regardless of name
            logo_f = self.photo_logo_path.get()
            
            # If path is empty, try smart fallback only if it hasn't been explicitly cleared
            # (In Photo Log, we usually want SITCO as default if nothing is selected)
            if not logo_f or not os.path.exists(logo_f):
                # Only fallback if the user hasn't explicitly cleared it in this session
                # For now, if empty, we try smart find for convenience
                if not logo_f:
                    logo_f = self.find_image_smart("SITCO")
                    if not logo_f: logo_f = self.find_image_smart("서울검사")
            
            if logo_f and os.path.exists(logo_f):
                try:
                    total_header_h = 45 
                    for r in range(1, 4): worksheet.set_row(r, 15) 
                    with PILImage.open(logo_f) as img:
                        w, h = img.size
                        max_w_logo = float(self.photo_logo_width_var.get())
                        mx = float(self.photo_logo_x_var.get())
                        my = float(self.photo_logo_y_var.get())
                        scale = min(max_w_logo/w, 42/h) * 0.95
                        logo_h = h * scale
                        y_offset = (total_header_h - logo_h) / 2 + my
                        worksheet.insert_image('A2', logo_f, {'x_scale': scale, 'y_scale': scale, 'x_offset': mx, 'y_offset': y_offset, 'object_position': 1})
                except Exception as e:
                    self.log(f"[PhotoLog] 로고 삽입 중 오류: {e}")

            worksheet.merge_range(1, 3, 1, 5, f"발주처: {self.photo_orderer.get()}", center_border)
            worksheet.merge_range(2, 3, 2, 5, f"REPORT NO: {self.photo_report_no.get()}", center_border)
            worksheet.merge_range(3, 3, 3, 5, f"검사일자: {self.photo_inspect_date.get()}", center_border)
            
            worksheet.set_row(4, 25)
            worksheet.merge_range(4, 0, 4, GRID_COLS-1, "PHOTO LOG (사진 대장)", bold_format)

            row = 5
            col_ptr = 0
            page_breaks = []
            DESC_ROW_HEIGHT = float(self.photo_desc_height_var.get())
            CELL_HEIGHT_PX = (CELL_ROW_HEIGHT * 1.33333) - 2
            
            total = len(image_files)
            current_row_max_h_pt = CELL_ROW_HEIGHT
            ROW_PT_TO_PX = 1.33333

            for i, img_path in enumerate(image_files):
                if col_ptr == 0:
                    current_row_max_h_pt = CELL_ROW_HEIGHT
                
                try:
                    with PILImage.open(img_path) as img:
                        img = ImageOps.exif_transpose(img)
                        img_w, img_h = img.size
                        
                        if self.photo_auto_rotate_var.get() and img_h > img_w:
                            img = img.rotate(90, expand=True)
                            img_w, img_h = img.size
                        
                        img_buffer = io.BytesIO()
                        img.save(img_buffer, format='PNG')
                        img_buffer.seek(0)
                        
                        c_start, c_end = photo_col_spans[col_ptr]
                        if c_start != c_end: worksheet.merge_range(row, c_start, row, c_end, "", center_border)
                        
                        if self.photo_fit_width_var.get():
                            x_scale = SAFE_WIDTH / img_w
                            y_scale = x_scale
                            req_h_px = (img_h * y_scale) + 10 
                            req_h_pt = req_h_px / ROW_PT_TO_PX
                            current_row_max_h_pt = max(current_row_max_h_pt, req_h_pt)
                        elif not self.photo_keep_aspect.get():
                            x_scale = SAFE_WIDTH / img_w
                            y_scale = (CELL_HEIGHT_PX - 10) / img_h
                        else:
                            scale = min(SAFE_WIDTH / img_w, (CELL_HEIGHT_PX - 10) / img_h)
                            x_scale = y_scale = scale
                        
                        worksheet.set_row(row, current_row_max_h_pt)
                        final_row_h_px = current_row_max_h_pt * ROW_PT_TO_PX
                        
                        x_off = round(((CELL_WIDTH_PX - (img_w * x_scale)) / 2) + SHIFT_X)
                        y_off = round(((final_row_h_px - (img_h * y_scale)) / 2) + SHIFT_Y)
                        
                        worksheet.insert_image(row, c_start, img_path, {
                            'image_data': img_buffer,
                            'x_scale': x_scale, 'y_scale': y_scale, 
                            'x_offset': x_off, 'y_offset': y_off, 
                            'object_position': 2 
                        })
                except Exception as e:
                    self.log(f"[PhotoLog Error] {os.path.basename(img_path)}: {e}")

                name = os.path.splitext(os.path.basename(img_path))[0]
                worksheet.set_row(row + 1, DESC_ROW_HEIGHT)
                c_start, c_end = photo_col_spans[col_ptr]
                worksheet.merge_range(row+1, c_start, row+1, c_end, f"사진 설명: {name}", desc_format)
                
                col_ptr += 1
                if col_ptr >= num_cols:
                    col_ptr = 0
                    row += 2
                
                if (i + 1) % photos_per_page == 0 and (i + 1) < total:
                    page_breaks.append(row if col_ptr==0 else row+2)

                self.progress["value"] = ((i + 1) / total) * 100
                self.status_var.set(f"사진대장 생성 중.. ({i+1}/{total})")

            if page_breaks: worksheet.set_h_pagebreaks(page_breaks)
            workbook.close()
            self.log(f"[PhotoLog] 완료: {os.path.basename(output_path)}")
            self.status_var.set("사진대장 생성 완료")
            messagebox.showinfo("성공", f"사진대장이 생성되었습니다.\n{output_path}")
            
            # [NEW] Open the folder
            try:
                os.startfile(os.path.dirname(output_path))
            except: pass

        except Exception as e:
            self.log(f"[PhotoLog Error] {e}")
            self.status_var.set("사진대장 생성 오류")
            messagebox.showerror("오류", f"작업 중 오류 발생:\n{e}")

    def setup_budget_tab(self, parent):
        container = tk.Frame(parent, background="#f9fafb"); container.pack(fill='both', expand=True, padx=10, pady=10)
        nb = ttk.Notebook(container); nb.pack(fill='both', expand=True)
        
        # 1. Labor Cost
        f_labor = ttk.Frame(nb); nb.add(f_labor, text=" 인건비 상세 ")
        self.labor_widget = LaborCostDetailWidget(f_labor, on_change_callback=lambda gt: self._on_budget_change()); self.labor_widget.pack(fill='both', expand=True)
        
        # 2. Material Cost
        f_mat = ttk.Frame(nb); nb.add(f_mat, text=" 재료비 상세 ")
        self.mat_cost_widget = MaterialCostDetailWidget(f_mat, on_change_callback=lambda gt: self._on_budget_change()); self.mat_cost_widget.pack(fill='both', expand=True)
        
        # 3. Expense & Profit
        f_exp = ttk.Frame(nb); nb.add(f_exp, text=" 경비 및 손익 ")
        self.exp_profit_widget = ExpenseProfitDetailWidget(f_exp, 
            get_labor_func=lambda: to_float(self.labor_widget.lbl_grand_total.cget('text').replace('₩','')),
            get_mat_func=lambda: to_float(self.mat_cost_widget.lbl_tot.cget('text').split('₩')[-1]),
            get_rev_func=lambda: 100000000 # Placeholder for revenue
        ); self.exp_profit_widget.pack(fill='both', expand=True)

    def _on_budget_change(self):
        if hasattr(self, 'exp_profit_widget'): self.exp_profit_widget.calculate_all()

    # --- UI & Data Helper Methods ---
    def populate_preview(self, data_list, switch_tab=True, mode="PMI"):
        """추출된 데이터를 미리보기 표에 채움 (필터 반영 및 그룹 색상 적용)"""
        if mode == "RT":
            tree = self.rt_preview_tree
            self.rt_item_idx_map = []
            idx_map = self.rt_item_idx_map
        elif mode == "PT":
            tree = self.pt_preview_tree
            self.pt_item_idx_map = []
            idx_map = self.pt_item_idx_map
        elif mode == "PAUT":
            tree = self.paut_preview_tree
            self.paut_item_idx_map = []
            idx_map = self.paut_item_idx_map
        else:
            tree = self.preview_tree
            self.item_idx_map = []
            idx_map = self.item_idx_map
            self.update_pmi_loc_listbox()
            
        for item in tree.get_children():
            tree.delete(item)
        
        filter_enabled = self.show_selected_only.get()
        hidden_by_def_count = 0
        total_matched_search = 0
        local_pmi_deficient_count = 0
        
        active_locs = set()
        if mode == "PMI" and hasattr(self, 'pmi_loc_listbox'):
            for i in range(self.pmi_loc_listbox.size()):
                val = self.pmi_loc_listbox.get(i)
                if val.startswith("[v] "): active_locs.add(val.replace("[v] ", "").lower())
        
        last_iso = None
        last_joint = None
        current_tag = "group_even"
        
        for idx, item in enumerate(data_list):
            if not item.get('date_filtered', True): continue
            
            is_selected = item.get('selected', True)
            if filter_enabled and not is_selected: continue
            
            if mode == "PMI":
                loc_val = str(item.get('Loc', '')).lower()
                if active_locs and loc_val not in active_locs: continue
                
                is_deficient = False
                if hasattr(self, 'element_filters'):
                    for f_item in self.element_filters:
                        f_key = f_item['key'].get().strip()
                        if not f_key: continue
                        f_min = round(self.to_float(f_item['min'].get()), 2)
                        f_max = round(self.to_float(f_item['max'].get()), 2)
                        val = round(self.to_float(item.get(f_key, item.get(f_key.capitalize(), 0.0))), 2)
                        if (f_min > 0 and val < f_min) or (f_max > 0 and val > f_max):
                            is_deficient = True; break
                
                if is_deficient: local_pmi_deficient_count += 1
                total_matched_search += 1
                
                if self.pmi_show_deficiency_only.get() and not is_deficient:
                    hidden_by_def_count += 1
                    continue
            else:
                is_deficient = False

            idx_map.append(idx)
            v_mark = "●" if is_selected else "○"
            st_mark = "⚠️" if is_deficient else "✅"
            
            curr_iso = item.get('Dwg', item.get('ISO', ''))
            norm_iso = self.normalize_iso(curr_iso)
            curr_joint = item.get('visual_group_joint', item.get('Joint', ''))
            
            is_new_iso = (last_iso is None or self.normalize_iso(last_iso) != norm_iso)
            is_new_joint = (last_joint is None or last_joint != curr_joint)
            
            if is_new_iso:
                current_tag = "group_odd" if current_tag == "group_even" else "group_even"
            
            display_count = len(idx_map) - 1 
            is_block_start = (display_count % 3 == 0)
            is_show = is_new_iso or is_new_joint or is_block_start

            last_iso = curr_iso
            last_joint = curr_joint
            
            row_vals = []
            if mode == "RT":
                for k in self.rt_column_keys:
                    if k == "selected": row_vals.append(v_mark)
                    elif k == "Acc":
                        val = item.get("Acc", "")
                        if not val:
                            res = str(item.get("Result", "")).upper()
                            val = "OK" if "ACC" in res or "OK" in res else ""
                        row_vals.append(val)
                    elif k == "Rej":
                        val = item.get("Rej", "")
                        if not val:
                            res = str(item.get("Result", "")).upper()
                            val = "NG" if "REJ" in res or "NG" in res or "RE" in res else ""
                        row_vals.append(val)
                    elif k == "Deg": row_vals.append(str(item.get("Deg", "")).strip())
                    else: 
                        val = str(item.get(k, "")).strip()
                        if (k in ["Dwg", "ISO"] and item.get('is_merged_iso')) or (k == "Joint" and item.get('is_merged_joint')): val = ""
                        row_vals.append(val)
            elif mode == "PT":
                for k in self.pt_column_keys:
                    if k == "selected": row_vals.append(v_mark)
                    else: 
                        val = str(item.get(k, "")).strip()
                        if (k in ["Dwg", "ISO"] and item.get('is_merged_iso')) or (k == "Joint" and item.get('is_merged_joint')): val = ""
                        row_vals.append(val)
            elif mode == "PAUT":
                for k in self.paut_column_keys:
                    if k == "No": row_vals.append(len(idx_map))
                    elif k == "selected": row_vals.append(v_mark)
                    else: 
                        val = str(item.get(k, "")).strip()
                        if (k in ["Dwg", "ISO"] and item.get('is_merged_iso')) or (k == "Joint" and item.get('is_merged_joint')): val = ""
                        row_vals.append(val)
            else: # PMI
                for k in self.column_keys:
                    if k == "_status": row_vals.append(st_mark)
                    elif k == "selected": row_vals.append(v_mark)
                    elif k in ["Ni", "Cr", "Mo", "Mn"]:
                        row_vals.append(f"{self.to_float(item.get(k, 0)):.2f}%")
                    else:
                        val = str(item.get(k, "")).strip()
                        if (k in ["Dwg", "ISO"] and item.get('is_merged_iso')) or (k == "Joint" and item.get('is_merged_joint')): val = ""
                        row_vals.append(val)

            row_tags = [str(idx), current_tag]
            if is_deficient: row_tags.append("deficient")
            tree.insert("", "end", values=tuple(row_vals), tags=tuple(row_tags))

    def update_date_listbox(self, mode="PMI"):
        """추출된 데이터에서 날짜 목록을 추출하여 리스트박스 갱신"""
        if mode == "RT": data = self.rt_extracted_data; lb = self.rt_date_listbox
        elif mode == "PT": data = self.pt_extracted_data; lb = self.pt_date_listbox
        elif mode == "PAUT": data = self.paut_extracted_data; lb = self.paut_date_listbox
        else: data = self.extracted_data; lb = self.pmi_date_listbox
        
        if not lb: return
        selected_dates = set()
        for i in range(lb.size()):
            val = lb.get(i)
            if val.startswith("[v] "): selected_dates.add(val.replace("[v] ", ""))
        
        unique_dates = sorted(list(set(str(item.get('Date', '')).strip() for item in data if item.get('Date'))))
        lb.delete(0, tk.END)
        for d in unique_dates:
            prefix = "[v] " if d in selected_dates else "[ ] "
            lb.insert(tk.END, prefix + d)

    def sort_by_column(self, col, mode="PMI"):
        """특정 컬럼 기준으로 데이터 정렬"""
        if mode == "RT": data = self.rt_extracted_data
        elif mode == "PT": data = self.pt_extracted_data
        elif mode == "PAUT": data = self.paut_extracted_data
        else: data = self.extracted_data
        
        if not data: return
        reverse = False
        if hasattr(self, '_last_sort_col') and self._last_sort_col == col:
            reverse = not getattr(self, '_last_sort_rev', False)
        
        self._last_sort_col = col
        self._last_sort_rev = reverse
        
        try:
            data.sort(key=lambda x: str(x.get(col, '')).lower(), reverse=reverse)
            self.populate_preview(data, switch_tab=False, mode=mode)
        except Exception as e:
            self.log(f"⚠️ 정렬 오류: {e}")

    def update_pmi_loc_listbox(self):
        if not hasattr(self, 'pmi_loc_listbox') or not self.extracted_data: return
        selected_locs = set()
        for i in range(self.pmi_loc_listbox.size()):
            val = self.pmi_loc_listbox.get(i)
            if val.startswith("[v] "): selected_locs.add(val.replace("[v] ", ""))
        current_locs = sorted(list(set(str(item.get('Loc', '')).strip() for item in self.extracted_data if item.get('Loc') and item.get('date_filtered', True))))
        self.pmi_loc_listbox.delete(0, tk.END)
        for loc in current_locs:
            prefix = "[v] " if loc in selected_locs else "[ ] "
            self.pmi_loc_listbox.insert(tk.END, prefix + loc)

    def normalize_iso(self, val):
        s_val = str(val).strip()
        try:
            if re.match(r'^\d+(\.\d+)?$', s_val):
                f_val = float(s_val)
                if f_val == int(f_val): return str(int(f_val))
                return str(f_val)
        except: pass
        return s_val.lower()

    def fix_material_name(self, t):
        t_str = str(t).upper()
        if t_str == 'NAN': return ""
        t_str = t_str.replace('A312-TP304', 'S/S').replace('A312-304L', 'S/S').replace('A312-305L', 'S/S').replace('A53-B', 'C/S').replace('A106-B', 'C/S')
        return t_str.replace('C2','C/S').replace('C4','C/S').replace('CS','C/S').replace('S99','S/S').replace('SS','S/S')

    def force_two_digit(self, val):
        try:
            s = str(val).strip()
            f = float(s)
            if f.is_integer(): return f"{int(f):02d}"
            return s
        except: return str(val).strip()

    def prepare_next_sheet(self, wb, source_sheet_idx, page_num):
        source_sheet = wb.worksheets[source_sheet_idx]
        new_sheet = wb.copy_worksheet(source_sheet) 
        
        # [FIX] openpyxl's copy_worksheet does not copy images. Manually copy them for RT to preserve Shooting Sketches.
        try:
            tab_idx = self.notebook.index("current")
            is_rt = (tab_idx == 5)
        except: is_rt = False

        if is_rt:
            try:
                import copy
                for img in source_sheet._images:
                    new_img = copy.copy(img)
                    if hasattr(img, 'anchor'):
                        new_img.anchor = copy.copy(img.anchor)
                    new_sheet.add_image(new_img)
            except: pass
            
        base_title = source_sheet.title.split('_')[0]
        new_sheet.title = f"{base_title[:20]}_{page_num:03d}"
        self.force_print_settings(new_sheet, context="DATA")
        
        # [FIX] Only add logos if they weren't already copied from the source sheet
        if not is_rt or not new_sheet._images:
            self.add_logos_to_sheet(new_sheet, is_cover=False)
        
        self.apply_custom_dimensions(new_sheet, "DATA")
        for col_letter, col_dim in source_sheet.column_dimensions.items(): 
            new_sheet.column_dimensions[col_letter].width = col_dim.width
            
        data_font = Font(size=9); grade_font = Font(size=8.5)
        
        # [FIX] RT 모드일 경우 Shooting Sketch(32행~42행)를 보호하기 위해 데이터 종료 행을 조절
        start_row = int(self.config.get('START_ROW', 17))
        end_row = int(self.config.get('DATA_END_ROW', 45))
        
        if is_rt:
            start_row = int(self.config.get('RT_START_ROW', 17))
            end_row = int(self.config.get('RT_END_ROW', 31))
            
        for r in range(start_row + 1, end_row + 1):
            for c in range(1, 14):
                cell = new_sheet.cell(row=r, column=c)
                cell.font = grade_font if c == 13 else data_font
                self.safe_set_value(new_sheet, cell, None)
                
        merged_to_clear = [rng for rng in new_sheet.merged_cells.ranges if rng.min_row >= start_row and rng.max_row <= end_row]
        for rng in merged_to_clear: new_sheet.unmerge_cells(str(rng))
        
        # [FIX] 갑지 데이터 수식으로 연결 (첫번째 시트 참조)
        try:
            ws0 = wb.worksheets[0]
            if len(wb.worksheets) > 1:
                self.safe_set_value(new_sheet, 'K5', f"='{ws0.title}'!L5")
                self.safe_set_value(new_sheet, 'M5', f"='{ws0.title}'!N5")
                self.safe_set_value(new_sheet, 'M8', f"='{ws0.title}'!N8")
                for r_idx in range(5, 11):
                    for c_idx in range(11, 14): # K(11) ~ M(13)
                        cell = new_sheet.cell(row=r_idx, column=c_idx)
                        cell.font = Font(name='바탕', size=9)
                        cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        except: pass
        return new_sheet

    def inject_drawing_layer(self, template_path, target_path):
        """[ADVANCED SURGERY] Multi-sheet drawing preservation logic."""
        import zipfile, shutil, os, re
        final_out = target_path + ".final"
        shutil.copy2(template_path, final_out)
        try:
            with zipfile.ZipFile(target_path, 'r') as z_data:
                ws_files = [f for f in z_data.namelist() if f.startswith('xl/worksheets/sheet') and f.endswith('.xml')]
                rep_shared_strings = []
                if 'xl/sharedStrings.xml' in z_data.namelist():
                    ss_content = z_data.read('xl/sharedStrings.xml').decode('utf-8', errors='ignore')
                    rep_shared_strings = re.findall(r'<t[^>]*>(.*?)</t>', ss_content, re.DOTALL)

                processed_sheets = {}
                for f in ws_files:
                    content = z_data.read(f).decode('utf-8', errors='ignore')
                    def replace_to_inline(match):
                        attrs, idx = match.group(1), int(match.group(2))
                        if 0 <= idx < len(rep_shared_strings):
                            text = rep_shared_strings[idx]; new_attrs = attrs.replace('t="s"', 't="inlineStr"')
                            return f'<c {new_attrs}><is><t>{text}</t></is>'
                        return match.group(0)
                    content = re.sub(r'<c ([^>]*t="s"[^>]*)>[ \t\r\n]*<v>(\d+)</v>', replace_to_inline, content)
                    processed_sheets[f] = content

                report_drawings = {f: z_data.read(f) for f in z_data.namelist() if 'xl/drawings/drawing' in f and f.endswith('.xml')}
                report_drawing_rels = {f: z_data.read(f) for f in z_data.namelist() if 'xl/drawings/_rels/drawing' in f}
                logo_media = {f: z_data.read(f) for f in z_data.namelist() if 'xl/media/image' in f}
                styles_xml = z_data.read('xl/styles.xml') if 'xl/styles.xml' in z_data.namelist() else None
                shared_strings_xml = z_data.read('xl/sharedStrings.xml') if 'xl/sharedStrings.xml' in z_data.namelist() else None
                workbook_xml = z_data.read('xl/workbook.xml') if 'xl/workbook.xml' in z_data.namelist() else None

                temp_swap = final_out + ".swap"
                with zipfile.ZipFile(final_out, 'r') as z_tmpl_in:
                    sheet_to_rid = {}
                    for f in z_tmpl_in.namelist():
                        if f.startswith('xl/worksheets/sheet') and f.endswith('.xml'):
                            content = z_tmpl_in.read(f).decode('utf-8', errors='ignore')
                            m = re.search(r'<drawing [^>]*r:id="(rId\d+)"', content)
                            if m: sheet_to_rid[f] = m.group(1)

                    with zipfile.ZipFile(temp_swap, 'w') as z_out:
                        for item in z_tmpl_in.infolist():
                            fname = item.filename
                            if fname in processed_sheets:
                                tmpl_sheet = z_tmpl_in.read(fname).decode('utf-8', errors='ignore')
                                # [PINCET SURGERY] Replace ONLY sheetData block to preserve all other tags (drawings, etc.)
                                data_match = re.search(r'<sheetData>(.*?)</sheetData>', processed_sheets[fname], re.DOTALL)
                                if data_match:
                                    new_sheet_xml = re.sub(r'<sheetData>.*?</sheetData>', f'<sheetData>{data_match.group(1)}</sheetData>', tmpl_sheet, flags=re.DOTALL)
                                    z_out.writestr(item, new_sheet_xml.encode('utf-8'))
                                else:
                                    z_out.writestr(item, tmpl_sheet.encode('utf-8'))
                            elif fname == 'xl/workbook.xml' and workbook_xml: z_out.writestr(item, workbook_xml)
                            elif fname == 'xl/styles.xml' and styles_xml: z_out.writestr(item, styles_xml)
                            elif fname == 'xl/sharedStrings.xml' and shared_strings_xml: z_out.writestr(item, shared_strings_xml)
                            else: z_out.writestr(item, z_tmpl_in.read(fname))
                        
                        # [NEW] Add extra sheets from report that are missing in template
                        # To keep print settings (Print Area, Page Breaks), we use the template's first sheet as a blueprint
                        blueprint_xml = None
                        first_sheet_name = next((f for f in z_tmpl_in.namelist() if f.startswith('xl/worksheets/sheet') and f.endswith('.xml')), None)
                        if first_sheet_name:
                            blueprint_xml = z_tmpl_in.read(first_sheet_name).decode('utf-8', errors='ignore')

                        tmpl_files = set(z_tmpl_in.namelist())
                        for f_report in processed_sheets:
                            if f_report not in tmpl_files:
                                # Use blueprint if available to inherit print settings/page breaks
                                if blueprint_xml:
                                    # [FIX] processed_sheets[f_report] is already a string in Final app (check previous turn)
                                    # Wait, let's check if it's bytes or string.
                                    # In Final app: processed_sheets[f] = content (string)
                                    report_sheet_content = processed_sheets[f_report]
                                    if isinstance(report_sheet_content, bytes):
                                        report_sheet_content = report_sheet_content.decode('utf-8')
                                    
                                    data_match = re.search(r'<sheetData>(.*?)</sheetData>', report_sheet_content, re.DOTALL)
                                    if data_match:
                                        extra_sheet_xml = re.sub(r'<sheetData>.*?</sheetData>', f'<sheetData>{data_match.group(1)}</sheetData>', blueprint_xml, flags=re.DOTALL)
                                        extra_sheet_xml = re.sub(r'<drawing [^>]*/>', '', extra_sheet_xml)
                                        z_out.writestr(f_report, extra_sheet_xml.encode('utf-8'))
                                    else:
                                        z_out.writestr(f_report, report_sheet_content.encode('utf-8'))
                                else:
                                    z_out.writestr(f_report, processed_sheets[f_report].encode('utf-8'))
                                
                                rel_path = f"xl/worksheets/_rels/{os.path.basename(f_report)}.rels"
                                if rel_path in z_data.namelist(): z_out.writestr(rel_path, z_data.read(rel_path))
                        tmpl_media_list = z_tmpl_in.namelist()
                        for f_path, f_data in logo_media.items():
                            if f_path not in tmpl_media_list: z_out.writestr(f_path, f_data)
            os.remove(target_path); os.rename(temp_swap, target_path)
            if os.path.exists(final_out): os.remove(final_out)
            self.log("   🛡️ [보호] 전 시트 제로-로스 데이터 주입 + 로고 복구 완료")
        except Exception as e:
            if os.path.exists(final_out): os.remove(final_out)
            self.log(f"   ⚠️ [주의] 제로-로스 수술 실패: {e}")

    def _write_gapji_metadata(self, ws):
        mapping = [
            ('GAPJI_PROJECT', 'B5'), ('GAPJI_CUSTOMER', 'B6'),
            ('GAPJI_ITEM', 'B7'), ('GAPJI_MATERIAL', 'B8'),
            ('GAPJI_EXAM_DATE', 'B9'), ('GAPJI_REPORT_NO', 'B10')
        ]
        for cfg_key, coord in mapping:
            val = self.config.get(cfg_key, "")
            if val: self.safe_set_value(ws, coord, val)

    def run_process(self):
        """성적서 생성 메인 오케스트레이터"""
        for k in list(self.config.keys()):
            if k.endswith(('_ROW', '_IDX', '_SIZE')) or any(x in k for x in ['START', 'END', 'PAGE']):
                try: self.config[k] = int(float(self.config[k]))
                except: pass
            elif any(x in k for x in ['MARGIN', 'SCALE', 'RATIO', 'POS']):
                try: self.config[k] = float(self.config[k])
                except: pass

        tab_idx = self.notebook.index("current")
        if tab_idx == 4: mode = "PMI"
        elif tab_idx == 5: mode = "RT"
        elif tab_idx == 6: mode = "PT"
        elif tab_idx == 7: mode = "PAUT"
        else:
            messagebox.showwarning("지원 안함", "해당 탭에서는 리포트 생성을 지원하지 않습니다.")
            return

        if mode == "RT":
            target_file = self.rt_target_file_path.get(); template_path = self.rt_template_file_path.get(); data = self.rt_extracted_data
        elif mode == "PT":
            target_file = self.pt_target_file_path.get(); template_path = self.pt_template_file_path.get(); data = self.pt_extracted_data
        elif mode == "PAUT":
            target_file = self.paut_target_file_path.get(); template_path = self.paut_template_file_path.get(); data = self.paut_extracted_data
        else: # PMI
            target_file = self.target_file_path.get(); template_path = self.template_file_path.get(); data = self.extracted_data
        
        if not template_path or not os.path.exists(template_path):
            messagebox.showwarning("오류", f"{mode} 양식 파일을 선택해주세요.")
            return
        
        # [NEW] Force capture all UI settings (Logo X/Y, Print Area, etc.) before generation
        self.manual_save_settings()
        if not data:
            if not self.extract_only(show_msg=False): return
            data = self.pt_extracted_data if mode == "PT" else (self.rt_extracted_data if mode == "RT" else (self.paut_extracted_data if mode == "PAUT" else self.extracted_data))
            
        final_list = [d for d in data if d.get('selected', True)]
        if not final_list:
            messagebox.showwarning("항목 미선택", f"선택된 {mode} 데이터가 없습니다.")
            return

        if mode == "RT": self._run_rt_process(final_list, template_path)
        elif mode == "PT": self._run_pt_process(final_list, template_path)
        elif mode == "PAUT": self._run_paut_process(final_list, template_path)
        else: self._run_pmi_process(final_list, template_path)

    def _run_pmi_process(self, final_list, template_path):
        """PMI 성적서 생성 (Block-based 레이아웃)"""
        self.log(f"🚀 PMI 성적서 생성 시작 (총 {len(final_list)} 건)...")
        self.progress['value'] = 0
        all_extracted_data = final_list
        data_start_row = int(self.config.get('START_ROW', 17))
        data_end_row = int(self.config.get('DATA_END_ROW', 45))
        
        try:
            wb = openpyxl.load_workbook(template_path)
            if len(wb.worksheets) < 1: raise ValueError("템플릿에 시트가 없습니다.")

            ws0 = wb.worksheets[0]; self.add_logos_to_sheet(ws0, is_cover=True)
            self.force_print_settings(ws0, context="COVER"); self.apply_custom_dimensions(ws0, "COVER")
            
            b_start = int(self.config.get('GAPJI_START_ROW', 23)); b_end = int(self.config.get('GAPJI_DATA_END_ROW', 38))
            if b_start > 0 and b_end >= b_start:
                for r in range(b_start, b_end + 1):
                    cell = ws0.cell(row=r, column=1); eb = cell.border
                    cell.border = Border(left=medium_side, right=eb.right, top=eb.top, bottom=eb.bottom)
            
            data_sheet_id = 1 if len(wb.worksheets) >= 2 else 0
            ws = wb.worksheets[data_sheet_id]; ws.title = f"PMI_Report_001"
            self.add_logos_to_sheet(ws, is_cover=False); self.force_print_settings(ws, context="DATA"); self.set_eulji_headers(ws)
            
            try:
                if len(wb.worksheets) >= 2:
                    self.safe_set_value(ws, 'K5', f"='{ws0.title}'!L5")
                    self.safe_set_value(ws, 'M5', f"='{ws0.title}'!N5")
                    self.safe_set_value(ws, 'M8', f"='{ws0.title}'!N8")
                    for r_idx in range(5, 11):
                        for c_idx in range(11, 14):
                            cell = ws.cell(row=r_idx, column=c_idx)
                            cell.font = Font(name='바탕', size=9); cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
            except: pass
            
            materials = "SS304,SS304L,SS316,SS316L,SS321,SS347,SS410,SS430,DUPLEX,MONEL,INCONEL,ER308,ER308L,ER309,ER309L,ER316,ER316L,ER347,ER2209,WP316,WP316L,TP316,TP316L,F316L,A182-F316L,A312-TP316L"
            dv_q = DataValidation(type="list", formula1=f'"{materials}"', allow_blank=True)
            ws.add_data_validation(dv_q)

            ni_col = self.col_to_num(self.config.get('PMI_COL_NI', '8')); cr_col = self.col_to_num(self.config.get('PMI_COL_CR', '9')); mo_col = self.col_to_num(self.config.get('PMI_COL_MO', '10'))
            print_elements = [('Ni', ni_col), ('Cr', cr_col), ('Mo', mo_col)]
            
            current_row = data_start_row; current_page = 1; data_ptr = 0
            while data_ptr < len(all_extracted_data):
                rows_left = data_end_row - current_row + 1
                if rows_left <= 0:
                    current_page += 1; ws = self.prepare_next_sheet(wb, data_sheet_id, current_page)
                    current_row = data_start_row; ws.add_data_validation(dv_q); rows_left = data_end_row - current_row + 1

                logical_block_size = 1; base_item = all_extracted_data[data_ptr]
                for ahead in range(data_ptr + 1, min(len(all_extracted_data), data_ptr + rows_left)):
                    if all_extracted_data[ahead].get('Joint') == base_item.get('Joint') and all_extracted_data[ahead].get('Dwg') == base_item.get('Dwg'):
                        logical_block_size += 1
                    else: break
                
                this_block_size = logical_block_size; batch = all_extracted_data[data_ptr : data_ptr + this_block_size]
                d_height = float(self.config.get('ROW_HEIGHT_DATA', 20.55))
                for r_off in range(this_block_size):
                    r = current_row + r_off
                    rd = ws.row_dimensions[r]; rd.height = d_height
                    for c in range(1, 14):
                        cell = ws.cell(row=r, column=c); l_s = thin_side; r_s = thin_side; t_s = thin_side; b_s = thin_side
                        if r == data_start_row: t_s = medium_side
                        if r == data_end_row: b_s = medium_side
                        if c == 1: l_s = medium_side
                        if c == 13: r_s = medium_side
                        cell.border = Border(left=l_s, right=r_s, top=t_s, bottom=b_s)

                self.safe_merge_cells(ws, current_row, 1, current_row + this_block_size - 1, 5)
                if this_block_size > 1: self.safe_merge_cells(ws, current_row, 6, current_row + this_block_size - 1, 6)
                
                if batch:
                    self.safe_set_value(ws, ws.cell(row=current_row, column=1).coordinate, batch[0].get('Dwg', ''))
                    self.safe_set_value(ws, ws.cell(row=current_row, column=6).coordinate, batch[0].get('Joint', ''))

                for i, item in enumerate(batch):
                    r = current_row + i
                    self.safe_set_value(ws, ws.cell(row=r, column=7).coordinate, item.get('Loc', ''), align='center')
                    for val_key, col_idx in print_elements:
                        v = self.to_float(item.get(val_key, 0.0))
                        cell = ws.cell(row=r, column=col_idx); self.safe_set_value(ws, cell.coordinate, v if v > 0 else "", align='center')
                        if v > 0: cell.number_format = '0.00'
                    self.safe_set_value(ws, ws.cell(row=r, column=13).coordinate, item.get('Grade', ''), align='center')
                    dv_q.add(ws.cell(row=r, column=13))

                data_ptr += len(batch); current_row += this_block_size
            
            output_file = filedialog.asksaveasfilename(defaultextension=".xlsx")
            if output_file: wb.save(output_file); messagebox.showinfo("성공", "PMI 성적서 생성 완료")
        except Exception as e: self.log(f"❌ PMI 오류: {e}")

    def _run_rt_process(self, final_list, template_path):
        """RT 성적서 생성 (High-Fidelity)"""
        self.manual_save_settings() 
        self.log(f"🚀 RT 성적서 생성 시작 (총 {len(final_list)} 건)...")
        self.progress['value'] = 0
        try:
            wb = openpyxl.load_workbook(template_path)
            if len(wb.worksheets) < 1: raise ValueError("템플릿에 시트가 없습니다.")

            ws0 = wb.worksheets[0]
            self.add_logos_to_sheet(ws0, is_cover=True, clear_existing=False)
            self._write_gapji_metadata(ws0)
            self.force_print_settings(ws0, context="COVER")
            self.apply_custom_dimensions(ws0, "COVER")
            
            data_sheet_id = 1 if len(wb.worksheets) >= 2 else 0
            ws = wb.worksheets[data_sheet_id]
            ws.title = f"RT_Report_001"
            self.add_logos_to_sheet(ws, is_cover=False, clear_existing=(ws != ws0))
            self.force_print_settings(ws, context="DATA")
            
            # RT Boundaries (Standard: Start 17, End 31 to protect Shooting Sketch at 32-42)
            start_row = int(self.config.get('RT_START_ROW', 17))
            end_row = int(self.config.get('RT_END_ROW', 31))
            current_row = start_row; current_page = 1; data_ptr = 0
            
            while data_ptr < len(final_list):
                if current_row > end_row:
                    current_page += 1
                    ws = self.prepare_next_sheet(wb, data_sheet_id, current_page)
                    current_row = start_row
                
                item = final_list[data_ptr]
                # Dynamic column mapping
                def _c(key, default): return self.col_to_num(self.config.get(key, default))
                
                self.safe_set_value(ws, ws.cell(row=current_row, column=_c('RT_COL_NO', '1')).coordinate, item.get('No', ''))
                self.safe_set_value(ws, ws.cell(row=current_row, column=_c('RT_COL_DWG', '3')).coordinate, item.get('Dwg', ''))
                self.safe_set_value(ws, ws.cell(row=current_row, column=_c('RT_COL_JOINT', '4')).coordinate, item.get('Joint', ''))
                
                # Defect columns (D1-D15)
                for d_i in range(1, 16):
                    d_col = _c(f'RT_COL_D{d_i}', str(12+d_i))
                    if d_col >= 1:
                        self.safe_set_value(ws, ws.cell(row=current_row, column=d_col).coordinate, item.get(f'D{d_i}', ''))

                self.safe_set_value(ws, ws.cell(row=current_row, column=_c('RT_COL_RES', '28')).coordinate, item.get('Result', 'ACC'))
                
                data_ptr += 1; current_row += 1
                self.progress['value'] = (data_ptr / len(final_list)) * 95

            # Save and surgery
            out = filedialog.asksaveasfilename(defaultextension=".xlsx")
            if out:
                wb.save(out)
                self.inject_drawing_layer(template_path, out)
                self.progress['value'] = 100
                messagebox.showinfo("성공", "RT 성적서 생성 완료 (도면 레이어 보호 적용)")
        except Exception as e:
            self.log(f"❌ RT 오류: {e}")
            messagebox.showerror("오류", str(e))

    def _run_pt_process(self, final_list, template_path):
        """PT 성적서 생성 (High-Fidelity)"""
        self.log(f"🚀 PT 성적서 생성 시작 (총 {len(final_list)} 건)...")
        try:
            wb = openpyxl.load_workbook(template_path)
            ws0 = wb.worksheets[0]; self.add_logos_to_sheet(ws0, is_cover=True)
            self.force_print_settings(ws0, context="COVER"); self.apply_custom_dimensions(ws0, "COVER")
            
            data_sheet_id = 1 if len(wb.worksheets) >= 2 else 0
            ws = wb.worksheets[data_sheet_id]; ws.title = f"PT_Report_001"
            self.add_logos_to_sheet(ws, is_cover=False); self.force_print_settings(ws, context="DATA")
            
            start_row = int(self.config.get('PT_START_ROW', 18)); end_row = int(self.config.get('PT_END_ROW', 37))
            current_row = start_row; current_page = 1; data_ptr = 0
            
            while data_ptr < len(final_list):
                if current_row > end_row:
                    current_page += 1; ws = self.prepare_next_sheet(wb, data_sheet_id, current_page); current_row = start_row
                
                item = final_list[data_ptr]
                self.safe_set_value(ws, ws.cell(row=current_row, column=1).coordinate, item.get('No', ''))
                self.safe_set_value(ws, ws.cell(row=current_row, column=2).coordinate, item.get('Dwg', ''))
                self.safe_set_value(ws, ws.cell(row=current_row, column=5).coordinate, item.get('Joint', ''))
                self.safe_set_value(ws, ws.cell(row=current_row, column=11).coordinate, item.get('Result', 'Acc'))
                
                for c in range(1, 12):
                    cell = ws.cell(row=current_row, column=c); cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True); cell.font = Font(name='바탕', size=9); cell.border = thin_border
                
                data_ptr += 1; current_row += 1
            
            out = filedialog.asksaveasfilename(defaultextension=".xlsx")
            if out: wb.save(out); messagebox.showinfo("성공", "PT 성적서 생성 완료")
        except Exception as e: self.log(f"❌ PT 오류: {e}")

    def _run_paut_process(self, final_list, template_path):
        self._generate_paut_report()

    def evaluate_paut_flaw(self, t, h, l, depth, nature, mode="ASME"):
        if 'crack' in str(nature).lower(): return "Reject", "Surface"
        return "Accept", "Subsurface"

    def _generate_paut_report(self):
        template_path = self.paut_template_file_path.get()
        final_list = [d for d in self.paut_extracted_data if d.get('selected', True)]
        if not template_path or not final_list: return
        try:
            wb = openpyxl.load_workbook(template_path); ws = wb.worksheets[0]
            start_row = int(self.config.get('PAUT_START_ROW', 11)); curr_row = start_row
            for i, item in enumerate(final_list):
                self.safe_set_value(ws, ws.cell(row=curr_row, column=2), item.get('Line No.', ''))
                self.safe_set_value(ws, ws.cell(row=curr_row, column=3), item.get('Joint No.', ''))
                self.safe_set_value(ws, ws.cell(row=curr_row, column=9), item.get('Evaluation', ''))
                curr_row += 1
            out = filedialog.asksaveasfilename(defaultextension=".xlsx")
            if out: wb.save(out); messagebox.showinfo("성공", "PAUT 성적서 생성 완료")
        except: pass

    def generate_photo_report(self):
        """사진 대장 생성 엔진 (xlsxwriter 기반)"""
        if not self.photo_selected_files: return
        try:
            output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="Photo_Report.xlsx")
            if not output_path: return
            workbook = xlsxwriter.Workbook(output_path); ws = workbook.add_worksheet()
            ws.set_paper(9); ws.set_margins(0.5, 0.5, 0.5, 0.5)
            fmt_title = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'border': 1})
            ws.merge_range(0, 0, 0, 5, self.photo_report_title.get(), fmt_title)
            
            cols = int(self.photo_cols_per_row.get()); row = 5; col_ptr = 0
            w_pct = float(self.photo_width_pct_var.get()) / 100.0
            cell_w_px = 350 if cols == 2 else 700
            for i, img_p in enumerate(self.photo_selected_files):
                with PILImage.open(img_p) as img:
                    img = ImageOps.exif_transpose(img)
                    w, h = img.size; scale = (cell_w_px * w_pct) / w
                    buf = io.BytesIO(); img.save(buf, format='PNG'); buf.seek(0)
                    ws.insert_image(row, col_ptr * (6//cols), img_p, {'image_data': buf, 'x_scale': scale, 'y_scale': scale, 'object_position': 2})
                    ws.write(row + 1, col_ptr * (6//cols), f"사진 설명: {os.path.basename(img_p)}")
                col_ptr += 1
                if col_ptr >= cols: col_ptr = 0; row += 2
            workbook.close(); messagebox.showinfo("성공", "사진대장 생성 완료")
        except Exception as e: messagebox.showerror("오류", str(e))


    def extract_only(self, show_msg=True):
        """데이터 추출 엔진 (PMI/RT/PT 통합)"""
        try:
            tab_idx = self.notebook.index("current")
            if tab_idx == 5: mode = "RT"
            elif tab_idx == 6: mode = "PT"
            elif tab_idx == 7: mode = "PAUT"
            else: mode = "PMI"
        except: mode = "PMI"

        if mode == "PAUT": return self._extract_paut_data()

        target_file = self.pt_target_file_path.get() if mode == "PT" else (self.rt_target_file_path.get() if mode == "RT" else self.target_file_path.get())
        if not target_file:
            messagebox.showwarning("파일 미선택", f"{mode} 데이터 파일을 선택해주세요.")
            return False
            
        self.log(f"🔍 {mode} 추출 시작: {os.path.basename(target_file)}")
        fname = os.path.basename(target_file)
        date_match = re.search(r'(\d{4}[-._]\d{2}[-._]\d{2}|\d{2}[-._]\d{2}[-._]\d{2}|\d{8}|\d{6})', fname)
        extracted_date = date_match.group(0) if date_match else datetime.datetime.now().strftime("%Y-%m-%d")
        
        all_extracted_data = []
        try:
            target_input = self.sequence_filter.get().strip()
            target_no_list = [x.strip() for x in target_input.replace(',', ' ').split() if target_input and x.strip()] if target_input else []
            
            with pd.ExcelFile(target_file) as xls:
                for s_idx, sheet_name in enumerate(xls.sheet_names):
                    try: temp_df = pd.read_excel(xls, sheet_name=sheet_name, header=None, nrows=50)
                    except: continue
                    header_idx = None
                    for i, row in temp_df.iterrows():
                        row_str = str(row.values).upper()
                        if mode == "RT" and ("FILM" in row_str or "DEFECT" in row_str or "IQI" in row_str): header_idx = i; break
                        elif mode == "PT" and ("LINE" in row_str and "JOINT" in row_str): header_idx = i; break
                        elif mode == "PMI" and ("CR" in row_str and "NI" in row_str): header_idx = i; break
                    if header_idx is None: continue
                    
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=header_idx)
                    if not df.empty and len(df) > 0:
                        row0_str = " ".join([str(x).upper() for x in df.iloc[0] if pd.notna(x)])
                        if any(k in row0_str for k in ["ORIGIN", "FACTOR", "WELDER1"]):
                            new_cols = []
                            for col, val in zip(df.columns, df.iloc[0]):
                                c_txt = str(col) if pd.notna(col) and "Unnamed" not in str(col) else ""
                                v_txt = str(val) if pd.notna(val) else ""
                                new_cols.append(f"{c_txt} {v_txt}".strip())
                            df.columns = new_cols; df = df.iloc[1:].reset_index(drop=True)

                    def _find_col(df, keywords, exclude=None):
                        for col in df.columns:
                            c_up = str(col).upper().strip()
                            if exclude and any(ex in c_up for ex in exclude): continue
                            if any(k == c_up for k in keywords): return col
                        for col in df.columns:
                            c_up = str(col).upper().strip()
                            if exclude and any(ex in c_up for ex in exclude): continue
                            if any(k in c_up for k in keywords):
                                if "NI" in keywords and ("UNIT" in c_up or "LINE" in c_up): continue
                                return col
                        return None

                    if mode == "RT":
                        col_no = _find_col(df, ["NO.", "SEQ"]); col_dwg = _find_col(df, ["ISO", "DWG", "LINE"])
                        col_joint = _find_col(df, ["JOINT", "J/N", "FILM IDENT"]); col_loc = _find_col(df, ["LOCATION", "POSITION"])
                        col_date = _find_col(df, ["DATE"]); col_acc = _find_col(df, ["ACC"]); col_rej = _find_col(df, ["REJ"])
                        col_result = _find_col(df, ["RESULT", "판정"]); defect_cols = {f"D{i}": _find_col(df, [f"D{i}", f"DEFECT{i}", chr(9311 + i)]) for i in range(1, 16)}
                    elif mode == "PT":
                        col_no = _find_col(df, ["NO."]); col_dwg = _find_col(df, ["ISO", "LINE"], exclude=["JOINT"])
                        col_joint = _find_col(df, ["JOINT"], exclude=["ISO", "LINE"]); col_mat = _find_col(df, ["MAT", "재질"])
                        col_size = _find_col(df, ["SIZE", "NPS"]); col_thk = _find_col(df, ["THK", "SCH"])
                        col_result = _find_col(df, ["RESULT", "판정"]); col_welder = _find_col(df, ["WELDER"])
                    else:
                        col_cr = _find_col(df, ["CR"]); col_ni = _find_col(df, ["NI"]); col_mo = _find_col(df, ["MO"])
                        col_no = _find_col(df, ["NO."]); col_joint = _find_col(df, ["JOINT"]); col_loc = _find_col(df, ["LOCATION"])
                        col_dwg = _find_col(df, ["ISO", "DWG"]); col_grade_orig = _find_col(df, ["GRADE", "MATERIAL"])

                    last_dwg = ""; last_joint = ""
                    for _, row in df.iterrows():
                        v_raw_no = str(row[col_no]).strip() if col_no is not None else str(_+1)
                        if target_no_list and v_raw_no not in target_no_list: continue
                        curr_dwg = str(row[col_dwg]).strip() if col_dwg is not None else ""
                        if (not curr_dwg or curr_dwg == "nan") and last_dwg: curr_dwg = last_dwg
                        if curr_dwg and curr_dwg != "nan": last_dwg = curr_dwg
                        curr_joint = str(row[col_joint]).strip() if col_joint is not None else ""
                        if (not curr_joint or curr_joint == "nan") and last_joint: curr_joint = last_joint
                        if curr_joint and curr_joint != "nan": last_joint = curr_joint
                        elif not curr_joint or curr_joint == "nan": curr_joint = v_raw_no

                        if mode == "RT":
                            it = {'No': v_raw_no, 'Date': str(row[col_date]).strip() if col_date else extracted_date, 'Dwg': curr_dwg, 'Joint': curr_joint, 'Loc': str(row[col_loc]).strip() if col_loc else "", 'Result': str(row[col_result]).strip() if col_result else "ACC", 'selected': True}
                            for i in range(1, 16):
                                v = str(row[defect_cols[f"D{i}"]]).strip() if defect_cols[f"D{i}"] else ""
                                it[f"D{i}"] = "√" if v.lower() in ["v", "x", "o", "1", "√"] else ""
                            all_extracted_data.append(it)
                        elif mode == "PT":
                            res_str = str(row[col_result]).upper() if col_result else "ACC"
                            if any(k in res_str for k in ["ACC", "OK", "PASS"]):
                                sz = str(row[col_size]).strip() if col_size else ""; tk = str(row[col_thk]).strip() if col_thk else ""
                                all_extracted_data.append({'No': v_raw_no, 'Date': extracted_date, 'Dwg': curr_dwg, 'Joint': self.force_two_digit(curr_joint), 'NPS': sz, 'Thk.': convert_sch_to_thk(sz, tk), 'Material': self.fix_material_name(row[col_mat]) if col_mat else "", 'Result': "Acc", 'selected': True})
                        else:
                            v_cr = self.to_float(row[col_cr]); v_ni = self.to_float(row[col_ni])
                            if v_cr > 0 or (v_raw_no != "" and v_raw_no != "nan"):
                                g = self.check_material_grade({'Cr': v_cr, 'Ni': v_ni, 'Mo': self.to_float(row[col_mo]) if col_mo else 0}) or "SS316"
                                all_extracted_data.append({'No': v_raw_no, 'Joint': curr_joint, 'Loc': str(row[col_loc]).strip() if col_loc else "", 'Cr': v_cr, 'Ni': v_ni, 'Mo': self.to_float(row[col_mo]) if col_mo else 0, 'Grade': g, 'Dwg': curr_dwg, 'Date': extracted_date, 'selected': True})
            
            if mode == "RT": self.rt_extracted_data.extend(all_extracted_data); self.sort_by_column("Dwg", mode="RT")
            elif mode == "PT": self.pt_extracted_data.extend(all_extracted_data); self.sort_by_column("Dwg", mode="PT")
            else: self.extracted_data.extend(all_extracted_data); self.sort_by_column("Dwg", mode="PMI")
            
            self.update_date_listbox(mode); self.progress['value'] = 100
            if show_msg: messagebox.showinfo("완료", f"{mode} 데이터 {len(all_extracted_data)}건 추출 완료")
            return True
        except Exception as e: self.log(f"❌ {mode} 추출 오류: {e}"); return False

    def _extract_paut_data(self):
        file_path = self.paut_target_file_path.get()
        if not file_path: return False
        try:
            df = pd.read_excel(file_path); cols = df.columns.tolist()
            mapping = {"t": next((c for c in cols if "THK" in c.upper() or "두께" in c.upper()), None),
                       "line": next((c for c in cols if "LINE" in c.upper() or "ISO" in c.upper()), None),
                       "joint": next((c for c in cols if "JOINT" in c.upper()), None)}
            self.paut_extracted_data = []
            for _, row in df.iterrows():
                self.paut_extracted_data.append({'selected': True, 'Line No.': str(row.get(mapping["line"], "")), 'Joint No.': str(row.get(mapping["joint"], "")), 'Th\'k(mm)': row.get(mapping["t"], 0), 'Evaluation': 'Accept'})
            self.populate_preview(self.paut_extracted_data, mode="PAUT"); return True
        except: return False

if __name__ == "__main__":
    root = tk.Tk()
    # Set app icon if available
    # try: root.iconbitmap(os.path.join(RESOURCE_DIR, "app_icon.ico"))
    # except: pass
    
    app = IntegratedSmartManager(root)
    root.mainloop()
