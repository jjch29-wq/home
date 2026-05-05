from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

path = os.path.abspath("../resources/Template_DailyWorkReport.xlsx")
wb = load_workbook(path)
sheet = wb.active

print("Headers in Row 42:")
for c in range(2, 20):
    val = sheet.cell(row=42, column=c).value
    if val:
        print(f"Col {get_column_letter(c)}: {val}")

wb.close()
