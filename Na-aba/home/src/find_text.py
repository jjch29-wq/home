from openpyxl import load_workbook
import os

path = os.path.abspath("../resources/Template_DailyWorkReport.xlsx")
wb = load_workbook(path)
sheet = wb.active

for r in range(1, 100):
    for c in range(1, 20):
        val = str(sheet.cell(row=r, column=c).value or '')
        if "300A" in val or "22P" in val:
            print(f"Found '{val}' at Row {r}, Col {c}")

wb.close()
