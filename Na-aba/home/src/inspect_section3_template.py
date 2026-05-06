from openpyxl import load_workbook
import os

path = os.path.abspath("../resources/Template_DailyWorkReport.xlsx")
wb = load_workbook(path)
sheet = wb.active

print(f"{'Row':<5} | {'A':<20} | {'B':<20} | {'E':<20}")
print("-" * 70)
for r in range(25, 35):
    a_val = sheet[f'A{r}'].value or ""
    b_val = sheet[f'B{r}'].value or ""
    e_val = sheet[f'E{r}'].value or ""
    
    # If cell is MergedCell, get value from top-left
    if not a_val:
        for rng in sheet.merged_cells.ranges:
            if r >= rng.min_row and r <= rng.max_row and 1 >= rng.min_col and 1 <= rng.max_col:
                a_val = f"(Merged {rng})"
                break
    
    if not b_val:
        for rng in sheet.merged_cells.ranges:
            if r >= rng.min_row and r <= rng.max_row and 2 >= rng.min_col and 2 <= rng.max_col:
                b_val = f"(Merged {rng})"
                break

    if not e_val:
        for rng in sheet.merged_cells.ranges:
            if r >= rng.min_row and r <= rng.max_row and 5 >= rng.min_col and 5 <= rng.max_col:
                e_val = f"(Merged {rng})"
                break

    print(f"{r:<5} | {str(a_val):<20} | {str(b_val):<20} | {str(e_val):<20}")

wb.close()
