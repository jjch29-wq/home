import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

output_path = r"C:\Users\-\OneDrive\바탕 화면\4.4.1_위험성평가표(RT)_완성본.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "위험성평가표"

# Styles
bold_font = Font(bold=True)
title_font = Font(bold=True, size=16)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
header_fill = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

def set_border(ws, min_col, min_row, max_col, max_row):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border

# Columns width setup
widths = {'A': 10, 'B': 10, 'C': 35, 'D': 30, 'E': 8, 'F': 8, 'G': 8, 'H': 35, 'I': 10, 'J': 10, 'K': 10, 'L': 8, 'M': 8, 'N': 8}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

# --- Header ---
ws.merge_cells('A1:N1')
ws['A1'] = "합동 위험성평가표 [ ■최초, □정기, □수시 ]"
ws['A1'].font = title_font
ws['A1'].alignment = center_align

# --- Table Headers ---
headers1 = ["세부작업", "유해 위험 요인 파악", "", "현재의 안전 보건 조치", "위험성 결정", "", "", "위험성 감소 대책", "개선\n예정일", "담당자\n(담당부서)", "완료일", "개선 후 위험성 (재평가)", "", ""]
headers2 = ["", "위험\n분류", "위험 발생 상황 및 결과", "", "가능성\n(빈도)", "중대성\n(강도)", "위험성\n(등급)", "", "", "", "", "가능성\n(빈도)", "중대성\n(강도)", "위험성\n(등급)"]

for col, val in enumerate(headers1, start=1):
    cell = ws.cell(row=3, column=col, value=val)
    cell.font = bold_font
    cell.alignment = center_align
    cell.fill = header_fill

for col, val in enumerate(headers2, start=1):
    cell = ws.cell(row=4, column=col, value=val)
    cell.font = bold_font
    cell.alignment = center_align
    cell.fill = header_fill

# Merge headers
ws.merge_cells('B3:C3')
ws.merge_cells('E3:G3')
ws.merge_cells('L3:N3')
ws.merge_cells('A3:A4')
ws.merge_cells('D3:D4')
ws.merge_cells('H3:H4')
ws.merge_cells('I3:I4')
ws.merge_cells('J3:J4')
ws.merge_cells('K3:K4')

set_border(ws, 1, 3, 14, 4)

# --- Data Rows ---
data = [
    ["필름부착 작업 시 배관상부에서 미끄러져 넘어짐 위험", "안전모, 안전화 착용 / 작업 전 점검", 3, 2, "V", "1. 2m 이상 고소작업 시 안전대(그네식) 체결 필수\n2. 우천/결빙 시 배관 상부 탑승 작업 중지"],
    ["지하 작업 중 낙하물에 의한 사고위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 굴착 상단부 자재 및 공구 적재 금지\n2. 하부 작업 시 상부 타 공정 동시 작업 금지"],
    ["필름부착작업 시 사다리 이동작업 중 추락/전도 위험", "안전모, 안전화 착용 / 작업 전 점검", 3, 2, "V", "1. A형 사다리 사용 시 2인 1조 작업 의무화\n2. 사다리 아웃트리거(전도방지대) 전개 및 최상단 탑승 금지"],
    ["차폐체 설치 작업 시 낙하에 의한 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 100kg 이상 중량물 인양 시 2줄 걸이 이상 결속\n2. 인양 중인 중량물 하부 통행 전면 금지"],
    ["차폐체설치 체인블럭 고정상태 불량시 낙하에 의한 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 체인블록 및 훅(Hook) 마모 상태 작업 전 점검\n2. 고정 지지대 하중 강도 확인 후 거치"],
    ["손상된 슬링벨트 사용 시 판단에 의한 중량물낙하 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 작업 전 슬링벨트 규격 및 손상 부위 확인\n2. 손상된 슬링벨트는 즉시 파기 및 교체"],
    ["중량물 줄걸이작업 부적합 시 낙하에 의한 사고위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 지정된 줄걸이 용구 사용 및 2줄 이상 결속\n2. 관리감독자(또는 신호수) 지휘 하 인양 진행"]
]

start_row = 5
for i, row_data in enumerate(data):
    r = start_row + i
    # 세부작업, 위험분류 (Merged for all rows)
    if i == 0:
        ws.cell(row=r, column=1, value="사전준비").alignment = center_align
        ws.cell(row=r, column=2, value="기계적\n(설비)\n요인").alignment = center_align
    
    # Data columns
    ws.cell(row=r, column=3, value=row_data[0]).alignment = left_align
    ws.cell(row=r, column=4, value=row_data[1]).alignment = left_align
    ws.cell(row=r, column=5, value=row_data[2]).alignment = center_align
    ws.cell(row=r, column=6, value=row_data[3]).alignment = center_align
    ws.cell(row=r, column=7, value=row_data[4]).alignment = center_align
    
    # 대책 column
    ws.cell(row=r, column=8, value=row_data[5]).alignment = left_align
    
    ws.row_dimensions[r].height = 50

ws.merge_cells(f'A5:A{start_row+len(data)-1}')
ws.merge_cells(f'B5:B{start_row+len(data)-1}')

set_border(ws, 1, 5, 14, start_row+len(data)-1)

wb.save(output_path)
print(f"Excel file created at: {output_path}")
