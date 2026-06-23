import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os

output_path = "4.4.1_위험성평가표(UT_표준양식).xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "위험성 평가서"

# Styles
bold_font = Font(bold=True)
title_font = Font(bold=True, size=20)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

def set_border(ws, min_col, min_row, max_col, max_row):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border

# Columns width setup (25 columns: A to Y)
# A~G: 11, 11, 30, 9, 5, 5, 6
# H~Y: 18 columns of width 6
widths = [11, 11, 30, 9, 5, 5, 6] + [6] * 18
for i, width in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

# --- Header Section (Rows 1 to 4) ---

# Row 1
ws['A1'] = "현 장 명"
ws.merge_cells('B1:D1')
ws['B1'] = "가산~가평 천연가스 공급시설 건설공사"

ws.merge_cells('E1:M2')
ws['E1'] = "위험성 평가서"
ws['E1'].font = title_font
ws['E1'].alignment = center_align

ws.merge_cells('E3:M3')
ws['E3'] = "[ ■ 최초평가   □ 수시평가   □ 정기평가 ]"
ws['E3'].font = bold_font
ws['E3'].alignment = Alignment(horizontal="center", vertical="center")

# Signatures Header
ws.merge_cells('N1:O1'); ws['N1'] = "작 성 자"
ws.merge_cells('P1:Q1'); ws['P1'] = "작성검토자"
ws.merge_cells('R1:W1'); ws['R1'] = "검 토 자"
ws.merge_cells('X1:Y1'); ws['X1'] = "승 인 자"

# Row 2
ws['A2'] = "작성일자"
ws.merge_cells('B2:D2')
ws['B2'] = "2026년 06월 15일"

ws.merge_cells('N2:O2'); ws['N2'] = "담당자"
ws.merge_cells('P2:Q2'); ws['P2'] = "공사담당자"
ws.merge_cells('R2:S2'); ws['R2'] = "보건관리자"
ws.merge_cells('T2:U2'); ws['T2'] = "안전관리자"
ws.merge_cells('V2:W2'); ws['V2'] = "안전관리팀장"
ws.merge_cells('X2:Y2'); ws['X2'] = "소 장"

# Row 3
ws['A3'] = "협력업체"
ws.merge_cells('B3:D3')
ws['B3'] = "서울검사(주)"

ws.merge_cells('N3:O3'); ws['N3'] = "(서명)"
ws.merge_cells('P3:Q3'); ws['P3'] = "(서명)"
ws.merge_cells('R3:S3'); ws['R3'] = "(서명)"
ws.merge_cells('T3:U3'); ws['T3'] = "(서명)"
ws.merge_cells('V3:W3'); ws['V3'] = "(서명)"
ws.merge_cells('X3:Y3'); ws['X3'] = "(서명)"

# Row 4
ws['A4'] = "대 공 종"
ws.merge_cells('B4:C4')
ws['B4'] = "초음파탐상검사"

ws.merge_cells('D4:E4'); ws['D4'] = "관리기간"
ws.merge_cells('F4:K4'); ws['F4'] = "2026년 06월 16일 ~ 2026년 06월 30일"

ws.merge_cells('L4:O4'); ws['L4'] = "현장소장\n승인의견"
ws.merge_cells('P4:Y4'); ws['P4'] = '"안전을 최우선으로"'

for r in range(1, 5):
    ws.row_dimensions[r].height = 25
    for c in range(1, 26):
        cell = ws.cell(row=r, column=c)
        cell.alignment = center_align
        cell.font = bold_font

ws.row_dimensions[3].height = 40 # 서명란 높이 확보
ws.row_dimensions[4].height = 30

set_border(ws, 1, 1, 25, 4)

# --- Table Headers (Rows 5 & 6) ---
ws.merge_cells('A5:A6'); ws['A5'] = "중공종"
ws.merge_cells('B5:B6'); ws['B5'] = "작업위치"
ws.merge_cells('C5:C6'); ws['C5'] = "위험성평가"
ws.merge_cells('D5:D6'); ws['D5'] = "재해\n형태"

ws.merge_cells('E5:G5'); ws['E5'] = "위험도"
ws['E6'] = "빈도"
ws['F6'] = "강도"
ws['G6'] = "등급\n(A,B,C)"

ws.merge_cells('H5:H6'); ws['H5'] = "중점\n등록\n(O)"
ws.merge_cells('I5:O6'); ws['I5'] = "위험성평가 개선대책"

ws.merge_cells('P5:Q5'); ws['P5'] = "작업일정"
ws.merge_cells('P6:Q6'); ws['P6'] = "작업인원"

ws.merge_cells('R5:S5'); ws['R5'] = "개선일정"
ws.merge_cells('R6:S6'); ws['R6'] = "개선책임자"

ws.merge_cells('T5:Y5'); ws['T5'] = "검토의견"
ws.merge_cells('T6:V6'); ws['T6'] = "공사담당\n공사팀장"
ws.merge_cells('W6:Y6'); ws['W6'] = "안전관리자\n보건관리자"

for r in range(5, 7):
    ws.row_dimensions[r].height = 25
    for c in range(1, 26):
        cell = ws.cell(row=r, column=c)
        cell.alignment = center_align
        cell.font = bold_font
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

set_border(ws, 1, 5, 25, 6)

# --- Data Definition ---
data = [
    # 1. 사전준비
    ["사전준비", "지하 작업 중 낙하물에 의한 사고위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 굴착 상단부에 자재 및 공구 적재 금지\n2. 하부 작업 시 상부 타 공정 혼재 작업 원칙적 금지"],
    ["사전준비", "손상된 슬링벨트 사용 시 판단에 의한 중량물낙하 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 작업 전 슬링벨트 손상 및 규격 확인\n2. 파손된 슬링벨트는 현장에서 즉시 파기 및 교체"],
    ["사전준비", "중량물 줄걸이작업 부적합 시 낙하에 의한 사고위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 지정된 줄걸이 용구 사용 및 결속 상태 교차 점검\n2. 관리감독자(또는 신호수) 지휘 하에 안전하게 인양 진행"],
    ["사전준비", "근로자가 안전모 등 개인 보호구 미착용에 의한 충돌", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 작업구역 진입 전 개인보호구 착용 상태 점검\n2. 안전모 턱끈 결속 철저"],
    ["사전준비", "중장비 이용 중량물 취급 작업 중 신호오류에 의한 충돌사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 장비 작업 반경 내 출입 통제 및 펜스 설치\n2. 전담 신호수 배치 및 운전원 간 무전 소통 확립"],

    # 2. 비파괴검사
    ["비파괴검사", "협소공간 검사 시 부자연스런 자세에 의한 근골격계질환 위험", "작업 전 스트레칭", 2, 1, "VI", "1. 작업 전·후 전신 스트레칭 실시\n2. 장시간 구부린 자세 유지 시 주기적인 휴식(10분 이상) 부여"],
    ["비파괴검사", "우천으로 인한 굴착법면상태 악화에 의한 붕괴 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 3, "V", "1. 작업 전 굴착 법면 균열, 용수(물비침) 현상 일일 점검\n2. 호우 경보 발효 시 굴착 지점 내부 진입 전면 금지"],

    # 3. 작업 후 정리
    ["작업 후 정리", "중량물 줄걸이작업 부적합 시 낙하에 의한 사고위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 탐상 장비 반출 시에도 2줄 걸이 이상 결속 원칙 준수\n2. 인양물 하부 접근 절대 금지"],
    ["작업 후 정리", "손상된 슬링벨트 사용 시 판단에 의한 중량물낙하 사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 철수 작업 전 슬링벨트 훼손 유무 재차 확인\n2. 손상 의심 시 무리한 재사용 금지 및 즉각 폐기"],
    ["작업 후 정리", "중장비 이용 중량물 취급 작업 중 신호오류에 의한 충돌사고 위험", "안전모, 안전화 착용 / 작업 전 점검", 2, 2, "V", "1. 장비 철수 동선 사전 파악 및 신호수/유도자 통제\n2. 운전원과 수신호 및 무전 교신 확인 후 이동"]
]

def extract_disaster_type(text):
    if "추락" in text: return "추락"
    if "낙하" in text: return "낙하"
    if "충돌" in text or "부딪" in text: return "충돌"
    if "넘어짐" in text or "미끄러" in text or "전도" in text: return "전도"
    if "근골격" in text: return "근골격계\n질환"
    if "누출" in text or "화상" in text: return "화학물질\n노출"
    if "방사능" in text or "노출" in text or "피폭" in text: return "방사선\n피폭"
    if "붕괴" in text: return "붕괴"
    return "기타"

start_row = 7
current_row = 7

# Grouping mechanism to merge "중공종", "작업위치" cells
groups = {}
for item in data:
    cat = item[0]
    groups[cat] = groups.get(cat, 0) + 1

for item in data:
    freq = item[3]
    sev = item[4]
    score = freq * sev
    
    if score >= 6:
        grade = "A"
        priority = "O"
    elif score >= 3:
        grade = "B"
        priority = "O"
    else:
        grade = "C"
        priority = ""
        
    ws.cell(row=current_row, column=1, value=item[0]).alignment = center_align
    ws.cell(row=current_row, column=2, value="비파괴검사 구역\n및 야적장").alignment = center_align
    ws.cell(row=current_row, column=3, value=item[1]).alignment = left_align
    ws.cell(row=current_row, column=4, value=extract_disaster_type(item[1])).alignment = center_align
    ws.cell(row=current_row, column=5, value=freq).alignment = center_align
    ws.cell(row=current_row, column=6, value=sev).alignment = center_align
    ws.cell(row=current_row, column=7, value=grade).alignment = center_align
    ws.cell(row=current_row, column=8, value=priority).alignment = center_align
    
    # 조치사항과 대책을 합치지 않고 개선 대책만 깔끔하게 출력 (I:O merge)
    improvement_text = f"{item[6]}"
    ws.merge_cells(f'I{current_row}:O{current_row}')
    ws.cell(row=current_row, column=9, value=improvement_text).alignment = left_align
    
    # 일정 및 책임자
    ws.merge_cells(f'P{current_row}:Q{current_row}')
    ws.cell(row=current_row, column=16, value="2026.06.16 -\n2026.06.30\n\n10명").alignment = center_align
    
    ws.merge_cells(f'R{current_row}:S{current_row}')
    ws.cell(row=current_row, column=18, value="2026.06.16 -\n\n관리감독자").alignment = center_align
    
    # 검토의견 빈칸
    ws.merge_cells(f'T{current_row}:V{current_row}')
    ws.merge_cells(f'W{current_row}:Y{current_row}')
    
    ws.row_dimensions[current_row].height = 65
    current_row += 1

# Merge '중공종' (Col A) and '작업위치' (Col B)
r = 7
for cat, count in groups.items():
    if count > 1:
        ws.merge_cells(f'A{r}:A{r+count-1}')
        ws.merge_cells(f'B{r}:B{r+count-1}')
    r += count

set_border(ws, 1, 7, 25, current_row - 1)

try:
    wb.save(output_path)
    print(f"Excel file created at: {output_path}")
except Exception as e:
    print(f"Error saving file: {e}")
