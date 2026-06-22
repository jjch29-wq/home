import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

output_path = r"C:\Users\-\OneDrive\바탕 화면\4.4.1_위험성평가표(RT)_완전체_전체공정.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "합동 위험성평가표"

# Styles
bold_font = Font(bold=True)
title_font = Font(bold=True, size=18)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

def set_border(ws, min_col, min_row, max_col, max_row):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border

# Columns width setup
widths = [10, 8, 8, 8, 6, 6, 9, 9, 9, 6, 6, 9, 9, 9, 9, 9]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

# --- Header Section (Rows 1 to 9) ---
ws.merge_cells('A1:P1')
ws['A1'] = "합동 위험성평가표 [ ■최초, □정기, □수시 ]"
ws['A1'].font = title_font
ws['A1'].alignment = center_align
ws.row_dimensions[1].height = 40

ws.merge_cells('A2:A3')
ws['A2'] = "작 업 명"
ws.merge_cells('B2:D3')
ws['B2'] = "가산~가평 천연가스 공급설비\n건설공사 비파괴검사 기술용역"
ws.merge_cells('E2:F2')
ws['E2'] = "평가기간"
ws.merge_cells('G2:I2')
ws['G2'] = "2026. 07. 09"
ws.merge_cells('J2:K2')
ws['J2'] = "재평가일"
ws.merge_cells('L2:N2')
ws.merge_cells('O2:P2')

ws.merge_cells('A4:A5')
ws['A4'] = "회 사 명"
ws.merge_cells('B4:D5')
ws['B4'] = "서울검사(주)"

ws.merge_cells('A6:A7')
ws['A6'] = "작업공종"
ws.merge_cells('B6:D7')
ws['B6'] = "방사선투과검사"

ws.merge_cells('A8:A9')
ws['A8'] = "작업기간"
ws.merge_cells('B8:D9')
ws['B8'] = "2026.06.22 ~ 2028.10.31"

ws.merge_cells('E3:E9')
ws['E3'] = "작업전"
ws.merge_cells('F3:F5')
ws['F3'] = "수급인"
ws['G3'] = "근 로 자"
ws['H3'] = "작 성 자"
ws['I3'] = "승 인 자"
ws.merge_cells('G4:G5')
ws['G4'] = "유상훈 (서명)"
ws.merge_cells('H4:H5')
ws['H4'] = "주진철 (서명)"
ws.merge_cells('I4:I5')
ws['I4'] = "강신태 (서명)"

ws.merge_cells('F6:F9')
ws['F6'] = "도급인"
ws['G6'] = "검토자(감독)"
ws['H6'] = "검토자(안전)"
ws['I6'] = "승 인 자"
ws.merge_cells('G7:G9')
ws['G7'] = "(서명)"
ws.merge_cells('H7:H9')
ws['H7'] = "(서명)"
ws.merge_cells('I7:I9')
ws['I7'] = "(서명)"

ws.merge_cells('J3:J9')
ws['J3'] = "재평가"
ws.merge_cells('K3:K5')
ws['K3'] = "수급인"
ws['L3'] = "근 로 자"
ws['M3'] = "작 성 자"
ws['N3'] = "승 인 자"
ws.merge_cells('L4:L5')
ws['L4'] = "(서명)"
ws.merge_cells('M4:M5')
ws['M4'] = "(서명)"
ws.merge_cells('N4:N5')
ws['N4'] = "(서명)"

ws.merge_cells('K6:K9')
ws['K6'] = "도급인"
ws['L6'] = "검토자(감독)"
ws['M6'] = "검토자(안전)"
ws['N6'] = "승 인 자"
ws.merge_cells('L7:L9')
ws['L7'] = "(서명)"
ws.merge_cells('M7:M9')
ws['M7'] = "(서명)"
ws.merge_cells('N7:N9')
ws['N7'] = "(서명)"

ws.merge_cells('O3:P9')

for r in range(2, 10):
    for c in range(1, 15):
        cell = ws.cell(row=r, column=c)
        cell.alignment = center_align
        cell.font = Font(size=10)

set_border(ws, 1, 2, 14, 9)

# --- Table Sub-Header (Rows 10-12) ---
ws.merge_cells('A10:I10')
ws['A10'] = "작업 전 평가"
ws.merge_cells('J10:N10')
ws['J10'] = "재평가"

ws.merge_cells('A11:A12')
ws['A11'] = "세부작업"
ws.merge_cells('B11:D11')
ws['B11'] = "유해 위험 요인 파악"
ws.merge_cells('E11:E12')
ws['E11'] = "현재의 안전 보건 조치"
ws.merge_cells('F11:H11')
ws['F11'] = "위험성 결정"
ws.merge_cells('I11:I12')
ws['I11'] = "위험성 감소 대책"
ws.merge_cells('J11:J12')
ws['J11'] = "개선\n예정일"
ws.merge_cells('K11:K12')
ws['K11'] = "담당자\n(담당부서)"
ws.merge_cells('L11:L12')
ws['L11'] = "완료일"
ws.merge_cells('M11:N11')
ws['M11'] = "개선 후 위험성 (재평가)"

ws['B12'] = "위험\n분류"
ws.merge_cells('C12:D12')
ws['C12'] = "위험 발생 상황 및 결과"
ws['F12'] = "가능성\n(빈도)"
ws['G12'] = "중대성\n(강도)"
ws['H12'] = "위험성\n(등급)"
ws['M12'] = "가능성\n(빈도)"
ws['N12'] = "중대성\n(강도)"

for r in range(10, 13):
    for c in range(1, 15):
        cell = ws.cell(row=r, column=c)
        cell.alignment = center_align
        cell.font = bold_font
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

set_border(ws, 1, 10, 14, 12)

# --- Data Definition ---
# Format: [세부작업, 위험분류, 위험발생상황, 안전보건조치, 빈도, 강도, 등급, 감소대책]
data = [
    # 1. 이동
    ["이동", "작업환경 요인", "이동 중 지면 단차, 돌출물, 자재 등에 걸려 넘어짐 (전도) 위험", "안전모, 안전화 착용 / 정리정돈", 2, 2, "IV", "1. 가설계단 및 이동통로 내 자재 정리정돈 철저\n2. 조도(밝기) 확보 및 보행 중 스마트폰 사용 금지 전파"],
    ["이동", "작업환경 요인", "이동용 사다리를 이용한 고소 이동 중 사다리 전도 및 추락", "안전모 착용 / 작업 전 TBM 전파", 3, 3, "V", "1. 사다리 상단 고정 및 아웃트리거 전개 확인\n2. 사다리 이용 시 반드시 2인 1조 작업"],
    ["이동", "작업특성 요인", "현장 내 중장비(굴착기 등) 주변 이동 중 장비와 충돌 또는 끼임", "작업 전 TBM 전파", 2, 3, "V", "1. 중장비 작업 반경 내 근로자 절대 출입 금지\n2. 전담 신호수 배치 및 운전원 무전기 소통 확립"],
    
    # 2. 사전준비
    ["사전준비", "기계적 요인", "필름부착 작업 시 배관상부에서 미끄러져 넘어짐 위험", "안전모, 안전화 착용 / 점검", 3, 2, "V", "1. 2m 이상 고소작업 시 지정된 생명줄에 안전대 체결 필수\n2. 우천/결빙 시 배관 상부 탑승 작업 전면 중지"],
    ["사전준비", "기계적 요인", "지하(트렌치) 작업 중 상부 낙하물에 의한 사고위험", "안전모, 안전화 착용", 2, 3, "V", "1. 굴착 상단부에 자재 및 공구 적재 금지\n2. 하부 작업 시 상부 타 공정 동시 작업 금지"],
    ["사전준비", "기계적 요인", "필름부착작업 시 사다리 작업 중 추락/전도 위험", "보호구 착용 / 점검", 3, 2, "V", "1. A형 사다리 최상단 발판 탑승 절대 금지\n2. 1인 하단부 지지 (2인 1조 작업)"],
    ["사전준비", "기계적 요인", "차폐체 설치/해체 작업 시 중량물 낙하에 의한 사고 위험", "보호구 착용 / 점검", 2, 3, "V", "1. 100kg 이상 인양 시 2줄 걸이 이상 결속\n2. 인양 중인 중량물 하부 통행 전면 금지"],
    ["사전준비", "기계적 요인", "인소스로봇(크롤러) 투입/반출 작업 시 낙하 위험", "작업 전 TBM 전파", 2, 3, "V", "1. 투입 전용 인양 장비 사용 및 결속 상태 2중 확인\n2. 배터리 및 조작계 이상 시 배관 내부 인력 투입 금지"],
    ["사전준비", "기계적 요인", "손상된 슬링벨트 사용 및 줄걸이 부적합으로 낙하 사고 위험", "안전모 착용 / 점검", 2, 3, "V", "1. 작업 전 슬링벨트 손상 규격 확인 및 파손품 폐기\n2. 관리감독자 지휘 하에 안전하게 인양 진행"],
    ["사전준비", "작업특성 요인", "협소공간 필름부착 시 부자연스런 자세에 의한 근골격계질환", "보호구 착용 / TBM 전파", 2, 2, "IV", "1. 작업 전·후 스트레칭 실시 및 중량물 인양 올바른 자세 교육\n2. 장시간 작업 시 교대 및 휴식 시간 부여"],
    ["사전준비", "작업환경 요인", "우천 등으로 인한 굴착법면 상태 약화로 붕괴 위험", "안전모 착용 / 점검", 2, 3, "V", "1. 작업 전 굴착 법면 균열, 용수(물비침) 점검\n2. 호우 경보 발효 시 굴착 지점 내부 진입 금지"],

    # 3. 비파괴검사
    ["비파괴검사", "화학적 요인", "필름 현상 작업 시 약품 취급 오류/누출에 의한 흡입 및 화상 위험", "보호구 착용 / 특별교육", 2, 3, "V", "1. 현상액/정착액 MSDS(물질안전보건자료) 비치 및 교육\n2. 작업 시 방독마스크, 내화학 장갑 착용 철저"],
    ["비파괴검사", "방사선 요인", "검사안전구역 미설정 및 통제 부실 시 방사능 노출 피폭 위험", "출입금지 간판 설치", 3, 3, "V", "1. 콜리메이터(차폐장비) 필수 장착으로 선량 최소화\n2. 방사선 10μSv/hr 이하 통제구역 설정 및 전담 감시자 배치\n3. 작업자 알람메타/개인선량계 항시 패용 (경보 시 대피)"],

    # 4. 작업 후 정리
    ["작업 후 정리", "기계적 요인", "필름 및 차폐체 해체 작업 시 배관상부 미끄러짐 및 추락/낙하", "보호구 착용 / TBM 전파", 3, 2, "V", "1. 정리 작업 중에도 안전대 체결 및 안전모 턱끈 결속 유지\n2. 해체된 자재는 던지지 않고 로프나 포대 등을 이용해 하강"],
    ["작업 후 정리", "기계적 요인", "웰딩하우스 하부 및 현장 정리 중 두부 충돌 및 부딪힘 위험", "보호구 착용 / TBM 전파", 2, 2, "IV", "1. 웰딩하우스 돌출 부위 야광 완충재 부착\n2. 철수 시에도 안전모 미착용 절대 금지"]
]

start_row = 13
current_row = 13

# Grouping mechanism to merge "세부작업" cells
groups = {}
for item in data:
    cat = item[0]
    groups[cat] = groups.get(cat, 0) + 1

for item in data:
    # item: [세부작업, 위험분류, 위험상황, 안전조치, 빈도, 강도, 등급, 대책]
    ws.cell(row=current_row, column=1, value=item[0]).alignment = center_align
    ws.cell(row=current_row, column=2, value=item[1]).alignment = center_align
    
    ws.merge_cells(f'C{current_row}:D{current_row}')
    ws.cell(row=current_row, column=3, value=item[2]).alignment = left_align
    ws.cell(row=current_row, column=5, value=item[3]).alignment = left_align
    ws.cell(row=current_row, column=6, value=item[4]).alignment = center_align
    ws.cell(row=current_row, column=7, value=item[5]).alignment = center_align
    ws.cell(row=current_row, column=8, value=item[6]).alignment = center_align
    
    ws.cell(row=current_row, column=9, value=item[7]).alignment = left_align
    
    ws.row_dimensions[current_row].height = 40
    current_row += 1

# Merge '세부작업' (Column A) blocks
r = 13
for cat, count in groups.items():
    if count > 1:
        ws.merge_cells(f'A{r}:A{r+count-1}')
    r += count

set_border(ws, 1, 13, 14, current_row - 1)

wb.save(output_path)
print(f"Excel file created at: {output_path}")
