import openpyxl
import re

# Header keyword mapping (keyword -> field name)
HEADER_MAP = {
    'Section': 'section',
    'Line No': 'line_no',
    'Line No.': 'line_no',
    'ORI': 'ori',
    'RE': 're',
    'TOTAL': 'total',
    '합격': 'result',
}

# Korean header keywords
KO_HEADER_MAP = {
    '업체': 'company',
    '번호': 'no',
    '구간': 'section',
    '라인번호': 'line_no',
    '관경': 'pipe_size',
    'Joint': 'joint',
    '용접사': 'welder',
    '결과': 'result',
    '검사성적': 'ori',
}

def find_paut_table(ws):
    """
    Scan the sheet to find the PAUT table header rows.
    Returns: (header_row, data_start_row, col_map)
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value.strip().replace('\n', ' ')
                if '1.2.1' in v or ('위상배열' in v and 'PAUT' in v):
                    # Found the section title - now scan below for header row
                    title_row = cell.row
                    print(f'Found 1.2.1 PAUT section at row {title_row}')
                    
                    # Scan rows below to find header
                    for hr in range(title_row + 1, title_row + 20):
                        row_cells = {}
                        has_section = False
                        has_lineno = False
                        for col in range(1, 30):
                            c = ws.cell(row=hr, column=col)
                            if c.value:
                                cv = str(c.value).strip().replace('\n', ' ')
                                if cv:
                                    row_cells[col] = cv
                                if 'Section' in cv:
                                    has_section = True
                                if 'Line No' in cv:
                                    has_lineno = True
                        
                        if has_section and has_lineno:
                            print(f'Found header row at {hr}: {row_cells}')
                            
                            # Check next row for sub-headers (ORI/RE/TOTAL)
                            sub_row = {}
                            for col in range(1, 30):
                                c = ws.cell(row=hr + 1, column=col)
                                if c.value:
                                    cv = str(c.value).strip().replace('\n', ' ')
                                    if cv:
                                        sub_row[col] = cv
                            print(f'Sub-header row {hr+1}: {sub_row}')
                            
                            # Build col_map
                            col_map = {}
                            
                            # From main header row
                            for col, val in row_cells.items():
                                if val.startswith('Section'):
                                    col_map['section'] = col
                                elif 'Line No' in val:
                                    col_map['line_no'] = col
                                elif 'PAUT' in val and '성적' in val:
                                    col_map['paut_start'] = col
                                elif val in ('업체', '회사'):
                                    col_map['company'] = col
                                elif val in ('번호',):
                                    col_map['no'] = col
                                elif val in ('관경', '구경'):
                                    col_map['pipe_size'] = col
                                elif 'Joint' in val:
                                    col_map['joint'] = col
                                elif '용접사' in val or '용접' in val:
                                    col_map['welder'] = col
                                elif '결과' in val:
                                    col_map['result'] = col
                            
                            # From sub-header row
                            for col, val in sub_row.items():
                                clean = val.replace("'", "").strip()
                                if clean == 'ORI':
                                    col_map['ori'] = col
                                elif clean == 'RE':
                                    col_map['re'] = col
                                elif 'TOTAL' in clean:
                                    col_map['total'] = col
                            
                            data_start = hr + 2
                            return hr, data_start, col_map
    
    return None, None, {}


# Test
wb = openpyxl.load_workbook(r'C:/Users/jjch2/Desktop/템플릿_최종완성본_V74.xlsx')
ws = wb.worksheets[0]
header_row, data_start, col_map = find_paut_table(ws)
print(f'\nHeader row: {header_row}')
print(f'Data starts at: {data_start}')
print(f'Column map: {col_map}')
