import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sys
import subprocess
import os
import time
import datetime
import json
import ctypes
import re
import traceback
import pandas as pd
import numpy as np

# Pre-compiled regex for performance
NAN_PATTERN = re.compile(r'^nan(\.0+)?$|^none$|^null$|^0\.0+|-0\.0+$', re.IGNORECASE)
DOT_ZERO_PATTERN = re.compile(r'\.0$')
MARKER_PATTERN = re.compile(r'\(.*?\)\s*|익일')

def install_and_import(package):
    try:
        __import__(package)
    except ImportError:
        try:
            print(f"Installing {package}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        except Exception as e:
            print(f"Failed to install {package}: {e}")
            messagebox.showerror("Dependency Error", f"Failed to install required package '{package}'.\nPlease install it manually: pip install {package}")
            sys.exit(1)

# Auto-install dependencies
try:
    import tkcalendar
    from tkcalendar import DateEntry, Calendar
    # [FIX] Patch tkcalendar for stability in ko_KR locale and prevent closure on navigation.
    _orig_cal_init = Calendar.__init__
    def _patched_cal_init(self, *args, **kwargs):
        # Default select_on_nav to False for better popup stability if not specified
        if 'select_on_nav' not in kwargs:
            kwargs['select_on_nav'] = False
        _orig_cal_init(self, *args, **kwargs)
        def widen():
            if hasattr(self, '_header_month'):
                try: self._header_month.configure(width=25)
                except: pass
            if hasattr(self, '_header_year'):
                try: self._header_year.configure(width=10)
                except: pass
        self.after_idle(widen)
        self._widen_headers = widen
    Calendar.__init__ = _patched_cal_init
    
    _orig_cal_display = Calendar._display_calendar
    def _patched_cal_display(self, *args, **kwargs):
        _orig_cal_display(self, *args, **kwargs)
        if hasattr(self, '_widen_headers'):
            self.after_idle(self._widen_headers)
    Calendar._display_calendar = _patched_cal_display

    _orig_de_show_cal = DateEntry._show_calendar
    def _patched_de_show_cal(self):
        try: self._last_known_date = self.get_date()
        except: self._last_known_date = None
        _orig_de_show_cal(self)
        if hasattr(self, '_calendar'):
            # Force select_on_nav=False to prevent closure during browsing
            try: self._calendar.configure(select_on_nav=False)
            except: pass
            self._calendar.after_idle(getattr(self._calendar, '_widen_headers', lambda: None))
    DateEntry._show_calendar = _patched_de_show_cal

    _orig_de_on_sel = DateEntry._on_calendar_selection
    def _patched_de_on_sel(self, event):
        # Only close if the date has changed vs what was there when opened, 
        # or if it was a real selection trigger.
        try:
            new_date = self._calendar.selection_get()
            if hasattr(self, '_last_known_date') and self._last_known_date == new_date:
                # Even if they re-click the same day, we might want to close.
                # But during navigation, select_on_nav=False ensures dates don't change.
                # If they clicked a day, selection_get() would be the clicked day.
                pass
        except: pass
        _orig_de_on_sel(self, event)
    DateEntry._on_calendar_selection = _patched_de_on_sel

    # [FIX] Python 3.13+ / 3.14 compat:
    # update_idletasks() 내부에서 KeyboardInterrupt(BaseException)가 전파될 수 있음.
    # 인스턴스 레벨에서 update_idletasks를 no-op으로 대체하여 스타일 설정 중 호출 자체를 차단.
    if hasattr(DateEntry, '_setup_style'):
        _orig_setup_style = DateEntry._setup_style
        def _patched_setup_style(self):
            _orig_update = self.update_idletasks
            self.update_idletasks = lambda: None  # 스타일 설정 중 update_idletasks 무시
            try:
                _orig_setup_style(self)
            except (KeyboardInterrupt, Exception):
                pass
            finally:
                self.update_idletasks = _orig_update
        DateEntry._setup_style = _patched_setup_style

except ImportError:
    install_and_import('tkcalendar')
    import tkcalendar
    from tkcalendar import DateEntry, Calendar
    # Re-apply patch after installation
    _orig_cal_init = Calendar.__init__
    def _patched_cal_init(self, *args, **kwargs):
        _orig_cal_init(self, *args, **kwargs)
        if hasattr(self, '_header_month'):
            try: self._header_month.configure(width=25)
            except: pass
    Calendar.__init__ = _patched_cal_init
except Exception:
    pass

# --- Custom Draggable Messagebox Implementation ---
class DraggableMessagebox:
    """A custom draggable replacements for standard tkinter.messagebox"""
    @staticmethod
    def _show(type, title, message):
        root = tk._default_root
        if not root:
            # Fallback to standard if no root exists yet
            if type == "error": return messagebox.showerror_orig(title, message)
            if type == "warning": return messagebox.showwarning_orig(title, message)
            return messagebox.showinfo_orig(title, message)

        dialog = tk.Toplevel(root)
        dialog.overrideredirect(True) # Remove standard title bar for better drag control
        dialog.attributes("-topmost", True)
        
        # Style the custom dialog
        dialog.config(background="#f3f4f6", highlightthickness=1, highlightbackground="#d1d5db")

        # Custom Title Bar
        title_bar = tk.Frame(dialog, background="#ffffff", height=30, cursor="fleur")
        title_bar.pack(side="top", fill="x")
        
        title_lbl = tk.Label(title_bar, text=title, font=("Malgun Gothic", 9, "bold"), background="#ffffff", padx=10)
        title_lbl.pack(side="left")
        
        def close_dialog():
            dialog.grab_release()
            dialog.destroy()

        btn_close = tk.Label(title_bar, text="✕", font=("Malgun Gothic", 10), background="#ffffff", padx=10, cursor="hand2")
        btn_close.pack(side="right")
        btn_close.bind("<Button-1>", lambda e: close_dialog())
        btn_close.bind("<Enter>", lambda e: btn_close.config(background="#ef4444", foreground="white"))
        btn_close.bind("<Leave>", lambda e: btn_close.config(background="#ffffff", foreground="black"))

        # Disable main window
        dialog.transient(root)
        dialog.grab_set()

        # Dragging logic
        def start_drag(event):
            dialog._drag_start_x = event.x
            dialog._drag_start_y = event.y

        def do_drag(event):
            x = dialog.winfo_x() + event.x - dialog._drag_start_x
            y = dialog.winfo_y() + event.y - dialog._drag_start_y
            dialog.geometry(f"+{x}+{y}")

        title_bar.bind("<Button-1>", start_drag)
        title_bar.bind("<B1-Motion>", do_drag)
        title_lbl.bind("<Button-1>", start_drag)
        title_lbl.bind("<B1-Motion>", do_drag)

        # Content
        main_frame = tk.Frame(dialog, background="#f3f4f6", padx=20, pady=20)
        main_frame.pack(expand=True, fill='both')

        icon_char = "ℹ" if type == "info" else "⚠" if type == "warning" else "❌"
        icon_color = "#0078d7" if type == "info" else "#f59e0b" if type == "warning" else "#ef4444"
        
        lbl_icon = tk.Label(main_frame, text=icon_char, font=("Malgun Gothic", 24), fg=icon_color, background="#f3f4f6")
        lbl_icon.pack(side="left", anchor="n", padx=(0, 15))

        # Use a wider wraplength for longer messages
        lbl_msg = tk.Label(main_frame, text=message, font=("Malgun Gothic", 10), justify="left", wraplength=480, background="#f3f4f6")
        lbl_msg.pack(side="left", fill="both", expand=True)

        # Auto-size: compute proper width/height after all widgets are built
        def _auto_position():
            dialog.update_idletasks()
            req_w = max(440, dialog.winfo_reqwidth() + 20)
            req_h = max(200, dialog.winfo_reqheight() + 20)
            x = root.winfo_x() + (root.winfo_width() // 2) - (req_w // 2)
            y = root.winfo_y() + (root.winfo_height() // 2) - (req_h // 2)
            dialog.geometry(f"{req_w}x{req_h}+{x}+{y}")
        dialog.after(10, _auto_position)

        btn_frame = tk.Frame(dialog, background="#f3f4f6", pady=10)
        btn_frame.pack(side="bottom", fill='x')
        
        # Standard tk.Button for crisp rectangular (사각형) look
        btn_ok = tk.Button(btn_frame, text="확인", font=("Malgun Gothic", 10, "bold"), 
                           command=close_dialog, background="#ffffff", activebackground="#e5e7eb",
                           relief="raised", borderwidth=1, padx=30, pady=5, width=10)
        btn_ok.pack(pady=10)
        btn_ok.focus_set()
        
        dialog.bind("<Return>", lambda e: close_dialog())
        dialog.bind("<Escape>", lambda e: close_dialog())

        dialog.lift()
        dialog.focus_force()
        root.wait_window(dialog)

    @staticmethod
    def showerror(title, message): DraggableMessagebox._show("error", title, message)
    @staticmethod
    def showwarning(title, message): DraggableMessagebox._show("warning", title, message)
    @staticmethod
    def showinfo(title, message): DraggableMessagebox._show("info", title, message)

# Injection: Replace standard messagebox methods to enable draggable behavior globally
if not hasattr(messagebox, 'showerror_orig'):
    messagebox.showerror_orig = messagebox.showerror
    messagebox.showwarning_orig = messagebox.showwarning
    messagebox.showinfo_orig = messagebox.showinfo
    
    messagebox.showerror = DraggableMessagebox.showerror
    messagebox.showwarning = DraggableMessagebox.showwarning
    messagebox.showinfo = DraggableMessagebox.showinfo





class SuggestionWindow:
    """A custom Toplevel window with a scrollable Label-based suggestion list."""
    def __init__(self, widget):
        self.widget = widget
        self.root = widget.winfo_toplevel()
        self.window = None
        self.canvas = None
        self.container = None
        self.labels = []
        self.active_index = -1
        self.active = False
        self.values = []
        self._drag_data = {"x": 0, "y": 0}
        self._just_selected = False # [NEW] Prevent re-show right after selection
        
        # Track parent window movement to reposition suggestions
        self.root.bind('<Configure>', lambda e: self._reposition(), add='+')

    def show(self, filtered_values):
        if not filtered_values or self._just_selected:
            self.hide()
            return

        self.values = filtered_values

        if not self.window or not self.window.winfo_exists():
            self.window = tk.Toplevel(self.root)
            self.window.withdraw()
            self.window.overrideredirect(True)
            self.window.attributes('-topmost', True)
            
            # Main border frame
            self.outer_frame = tk.Frame(self.window, background="#888888", padx=1, pady=1)
            self.outer_frame.pack(fill='both', expand=True)
            
            # Scrollbar and Canvas
            self.scrollbar = ttk.Scrollbar(self.outer_frame, orient="vertical")
            self.canvas = tk.Canvas(self.outer_frame, background="white", highlightthickness=0, yscrollcommand=self.scrollbar.set)
            
            self.container = tk.Frame(self.canvas, background="white")
            self.canvas_window = self.canvas.create_window((0, 0), window=self.container, anchor="nw")
            
            self.scrollbar.config(command=self.canvas.yview)
            
            self.canvas.pack(side="left", fill="both", expand=True)
            self.scrollbar.pack(side="right", fill="y")
            
            # Bind mousewheel to all components
            self.window.bind("<MouseWheel>", self._handle_mousewheel)
            self.outer_frame.bind("<MouseWheel>", self._handle_mousewheel)
            self.canvas.bind("<MouseWheel>", self._handle_mousewheel)
            self.container.bind("<MouseWheel>", self._handle_mousewheel)
            self.scrollbar.bind("<MouseWheel>", self._handle_mousewheel)
            
            # Linux/Unix mouse wheel bindings (Button-4, Button-5)
            self.window.bind("<Button-4>", self._handle_mousewheel)
            self.window.bind("<Button-5>", self._handle_mousewheel)
            self.canvas.bind("<Button-4>", self._handle_mousewheel)
            self.canvas.bind("<Button-5>", self._handle_mousewheel)
            self.container.bind("<Button-4>", self._handle_mousewheel)
            self.container.bind("<Button-5>", self._handle_mousewheel)
            
            # Make it draggable (via the outer frame/border)
            def start_drag(event):
                self._drag_data["x"] = event.x
                self._drag_data["y"] = event.y
                # Optional: lifting window when interaction starts
                self.window.lift()
            
            def do_drag(event):
                x = self.window.winfo_x() + (event.x - self._drag_data["x"])
                y = self.window.winfo_y() + (event.y - self._drag_data["y"])
                self.window.geometry(f"+{x}+{y}")
            
            self.outer_frame.bind("<Button-1>", start_drag)
            self.outer_frame.bind("<B1-Motion>", do_drag)

        # Clear existing labels
        for lbl in self.labels:
            lbl.destroy()
        self.labels = []
        self.active_index = -1

        # Create new labels for each value
        for i, val in enumerate(filtered_values):
            lbl = tk.Label(self.container, text=val, font=('Malgun Gothic', 10),
                          anchor='w', padx=5, pady=2, background="white", cursor="hand2")
            lbl.pack(fill='x')
            
            # Bind events to each label individually
            lbl.bind('<Enter>', lambda e, idx=i: self._highlight(idx))
            lbl.bind('<Button-1>', lambda e, v=val: self._select(v))
            lbl.bind('<MouseWheel>', self._handle_mousewheel)
            lbl.bind('<Button-4>', self._handle_mousewheel)
            lbl.bind('<Button-5>', self._handle_mousewheel)
            self.labels.append(lbl)
        
        # [NEW] Enhanced Width Calculation
        self.container.update_idletasks()
        widget_width = self.widget.winfo_width()
        
        # Calculate visual width based on content
        content_width = 0
        try:
            def get_visual_width(s):
                # Korean/Full-width chars are approx 2x wider than half-width
                return sum(1.8 if ord(c) > 127 else 1.0 for c in str(s))
            
            if filtered_values:
                max_visual_chars = max(get_visual_width(v) for v in filtered_values)
                # Appox 8-9px per half-width char for Malgun Gothic 10
                content_width = int(max_visual_chars * 9) + 25 
        except:
            content_width = 0
            
        # Use the larger of widget width or content width
        width = max(widget_width, content_width)
        
        item_height = 24
        max_items = 12
        visible_items = min(len(filtered_values), max_items)
        height = visible_items * item_height
        
        # Adjust container width in canvas
        self.canvas.itemconfig(self.canvas_window, width=width)
        self.canvas.config(scrollregion=(0, 0, width, len(filtered_values) * item_height))
        
        # Toggle scrollbar
        if len(filtered_values) > max_items:
            self.scrollbar.pack(side="right", fill="y")
        else:
            self.scrollbar.pack_forget()

        # Position exactly below the widget
        x = self.widget.winfo_rootx()
        y = self.widget.winfo_rooty() + self.widget.winfo_height()
        
        self.window.geometry(f"{int(width+20 if len(filtered_values) > max_items else width)}x{int(height)}+{int(x)}+{int(y)}")
        self.window.deiconify()
        self.window.lift()
        self.window.update_idletasks()
        self.active = True

    def _handle_mousewheel(self, event):
        """Unified mousewheel handler for all child widgets"""
        if self.active and self.canvas:
            if event.num == 4: # Linux scroll up
                self.canvas.yview_scroll(-1, "units")
            elif event.num == 5: # Linux scroll down
                self.canvas.yview_scroll(1, "units")
            else: # Windows/macOS delta
                self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            return "break"

    def _reposition(self):
        """Update window position to stay attached to the widget"""
        if self.active and self.window and self.window.winfo_exists() and self.window.winfo_viewable():
            x = self.widget.winfo_rootx()
            y = self.widget.winfo_rooty() + self.widget.winfo_height()
            # Preserve current width/height
            curr_geom = self.window.geometry().split('+')[0]
            self.window.geometry(f"{curr_geom}+{int(x)}+{int(y)}")

    def hide(self, event=None):
        self.active = False
        if self.window and self.window.winfo_exists():
            try:
                self.window.withdraw()
                # Use update() to force immediate visual removal
                self.window.update() 
            except: pass

    def _highlight(self, index):
        # Remove previous highlight
        if self.active_index >= 0 and self.active_index < len(self.labels):
            self.labels[self.active_index].config(background="white", foreground="black")
            
        self.active_index = index
        if self.active_index >= 0 and self.active_index < len(self.labels):
            lbl = self.labels[self.active_index]
            lbl.config(background="#0078d7", foreground="white")
            
            # Ensure visible in scroll context
            self._ensure_visible(index)

    def _ensure_visible(self, index):
        """Scroll canvas to make the highlighted label visible."""
        if not self.canvas: return
        
        item_height = 24
        total_items = len(self.labels)
        if total_items == 0: return
        
        # Fraction of the way down
        top = index / total_items
        bottom = (index + 1) / total_items
        
        # Current view range
        view_top, view_bottom = self.canvas.yview()
        
        if top < view_top:
            self.canvas.yview_moveto(top)
        elif bottom > view_bottom:
            self.canvas.yview_moveto(bottom - (view_bottom - view_top))

    def _select(self, val):
        self._just_selected = True
        self.hide() # Hide immediately before any delay or event generation
        def commit():
            try:
                self.widget.delete(0, tk.END)
                self.widget.insert(0, val)
                self.hide() # Double-ensure hidden
                # Trigger events for auto-save/search
                self.widget.event_generate('<<ComboboxSelected>>')
                # self.widget.event_generate('<Return>') # [NOTICE] Generating Return can re-trigger keyrelease filters
            except: pass
            finally:
                # Reset cooldown after a delay to allow UI to settle
                self.widget.after(500, lambda: setattr(self, '_just_selected', False))
        
        # Small visual feedback delay
        self.widget.after(20, commit)

    def move_selection(self, direction):
        if not self.active or not self.labels: return
        
        new_idx = (self.active_index + direction) % len(self.labels)
        self._highlight(new_idx)
        return True

    def confirm_selection(self):
        if not self.active: return False
        if self.active_index >= 0 and self.active_index < len(self.values):
            self._select(self.values[self.active_index])
            return True
        self.hide()
        return False


def register_autocomplete(combobox, suggestion_list):
    """Enable SuggestionWindow-based autocomplete for any ttk.Combobox"""
    # ── 이미 등록된 경우 이벤트 핸들러 중복 추가 방지 ──────────────────
    if hasattr(combobox, '_suggestion_win'):
        return  # combobox['values'] 갱신은 호출 측에서 직접 처리

    combobox._suggestion_win = SuggestionWindow(combobox)
    combobox._autocomplete_timer = None

    def perform_filter(force_all=False):
        typed = combobox.get()
        all_vals = combobox['values']
        if not all_vals: return
        
        # [FIX] If forced (click/focus), show all values even if something is typed/selected
        if force_all or not typed:
            combobox._suggestion_win.show(all_vals)
            return
        
        if len(typed) == 1:
            filtered = [v for v in all_vals if str(v).lower().startswith(typed.lower())]
        else:
            # Substring match for 2+ characters
            filtered = [v for v in all_vals if typed.lower() in str(v).lower()]
        
        combobox._suggestion_win.show(filtered)

    def on_keyrelease(event):
        if event.keysym == "Down" and combobox._suggestion_win.active:
            combobox._suggestion_win.move_selection(1)
            return "break"
        if event.keysym == "Up" and combobox._suggestion_win.active:
            combobox._suggestion_win.move_selection(-1)
            return "break"
        
        # [NOTICE] Removed "Return" handling here as it's already handled in handle_extra_nav (KeyPress).
        # Double handling Return (KeyPress + KeyRelease) causes focus skipping (double-jump) 
        # when moving focus between autocomplete widgets.
        
        if event.keysym in ("Left", "Right", "Up", "Down", "Return", "Escape", "Tab", "Shift_L", "Shift_R", "Control_L", "Control_R"):
            if event.keysym in ("Escape", "Tab"):
                combobox._suggestion_win.hide()
            return
        
        if hasattr(combobox, '_autocomplete_timer') and combobox._autocomplete_timer:
            combobox.after_cancel(combobox._autocomplete_timer)
        combobox._autocomplete_timer = combobox.after(50, lambda: perform_filter(force_all=False))
    
    combobox.bind('<KeyRelease>', on_keyrelease, add="+")
    
    def handle_extra_nav(event):
        if not combobox._suggestion_win.active: return
        if event.keysym == "Down": combobox._suggestion_win.move_selection(1); return "break"
        if event.keysym == "Up": combobox._suggestion_win.move_selection(-1); return "break"
        if event.keysym == "Return": 
            if combobox._suggestion_win.confirm_selection():
                # Trigger search if bound to the combobox
                combobox.event_generate('<<ComboboxSelected>>')
                return "break"
            
    combobox.bind('<KeyPress-Down>', handle_extra_nav, add="+")
    combobox.bind('<KeyPress-Up>', handle_extra_nav, add="+")
    combobox.bind('<KeyPress-Return>', handle_extra_nav, add="+")

    def on_focus_out(event):
        combobox.after(200, combobox._suggestion_win.hide)
    combobox.bind('<FocusOut>', on_focus_out, add=True)
    combobox.bind('<<ComboboxSelected>>', lambda e: combobox._suggestion_win.hide(), add=True)

    def on_focus_in(event):
        # [FIX] Trigger filter with force_all=True on focus
        perform_filter(force_all=True)
    combobox.bind('<FocusIn>', on_focus_in, add="+")
    
    # 텍스트 영역 클릭: SuggestionWindow 토글 (열려 있으면 닫기, 닫혀 있으면 열기)
    def _on_click(event):
        if combobox._suggestion_win.active:
            combobox._suggestion_win.hide()
        else:
            perform_filter(force_all=True)
    combobox.bind('<Button-1>', _on_click, add="+")

    # ── ▼ 버튼(postcommand): native 드롭다운 차단 + SuggestionWindow 토글 ──
    def _postcommand():
        _saved = list(combobox['values'])
        if combobox._suggestion_win.active:
            # 이미 열려 있으면 → 닫기
            combobox._suggestion_win.hide()
        else:
            # 닫혀 있으면 → 열기
            if _saved:
                combobox._suggestion_win.show(_saved)
        # values=[] 로 native 드롭다운 항상 차단 (Post가 값 없으면 건너뜀)
        combobox.configure(values=[])
        combobox.after_idle(lambda: combobox.configure(values=_saved))
    combobox.configure(postcommand=_postcommand)


class WorkerCompositeWidget(ttk.Frame):
    """
    Composite widget for Worker selection: [Name] with Autocomplete
    """
    def __init__(self, parent, enable_autocomplete=False, user_list=None, **kwargs):
        super().__init__(parent)
        
        # Worker Name selection
        name_width = kwargs.pop('width', 15)
        self.cb_name = ttk.Combobox(self, width=name_width, **kwargs)
        self.cb_name.pack(side='left', fill='x', expand=True)
        
        # Enable autocomplete if requested
        if enable_autocomplete and user_list is not None:
            self._autocomplete_timer = None
            register_autocomplete(self.cb_name, user_list)
        
    def get(self):
        """Return clean name"""
        return self.cb_name.get().strip()

    def set(self, value):
        """Set name, cleaning off any (Shift) prefixes if present"""
        if not value:
            self.cb_name.set("")
            return
            
        import re
        # Progressively migrate: if data still has (Shift) prefix, strip it for the name field
        match = re.match(r"\((주간|야간|휴일|주야간)\)\s*(.*)", str(value))
        if match:
            self.cb_name.set(match.group(2).strip())
        else:
            self.cb_name.set(str(value).strip())

    def bind(self, sequence=None, func=None, add=None):
        self.cb_name.bind(sequence, func, add)

    def current(self, newindex=None):
        return self.cb_name.current(newindex)
        
    def config(self, **kwargs):
        self.cb_name.config(**kwargs)

    def __setitem__(self, key, value):
        self.cb_name[key] = value

    def __getitem__(self, key):
        return self.cb_name[key]

class WorkerDataGroup(ttk.Frame):
    """
    Unified widget for a worker's record: [Name] [Shift] [WorkTime] [OT]
    """
    def __init__(self, parent, worker_index, users_list, time_list=None, enable_autocomplete=False, **kwargs):
        super().__init__(parent, padding=2) # Reduced padding for compact layout
        self.worker_index = worker_index
        
        # 1. Name selection (WorkerCompositeWidget now handles only name)
        self.composite = WorkerCompositeWidget(
            self, width=8, values=users_list, 
            enable_autocomplete=enable_autocomplete, 
            user_list=users_list
        )
        self.composite.pack(side='left', padx=(0, 2))
        self.cb_name = self.composite.cb_name
        
        # 2. Shift selection (Moved here from WorkerCompositeWidget)
        self.cb_shift = ttk.Combobox(self, values=["주간", "야간", "휴일", "주야간"], width=6, state="readonly")
        self.cb_shift.pack(side='left', padx=(1, 2))
        self.cb_shift.set("") # Default empty
        
        # 3. Work Time (Changed to Combobox for mouse selection)
        ttk.Label(self, text="시간:").pack(side='left', padx=(1, 0))
        self.ent_worktime = ttk.Combobox(self, width=16, values=time_list or [])
        self.ent_worktime.pack(side='left', padx=(0, 2))
        self.ent_worktime.set("") # Default empty
        
        # 4. OT
        ttk.Label(self, text="OT:").pack(side='left', padx=(1, 0))
        self.ent_ot = ttk.Entry(self, width=18)
        self.ent_ot.pack(side='left')

    def get_worker(self): return self.composite.get()
    def set_worker(self, val): self.composite.set(val)
    
    def get_time(self): 
        """Return combined string: (Shift) Time"""
        shift = self.cb_shift.get()
        time = self.ent_worktime.get().strip()
        if not time:
            return ""
        return f"({shift}) {time}"
        
    def set_time(self, val):
        """Parse string '(Shift) Time' and set widgets"""
        if not val:
            self.ent_worktime.set("")
            self.cb_shift.set("주간")
            return
            
        import re
        match = re.match(r"\((주간|야간|휴일|주야간)\)\s*(.*)", str(val))
        if match:
            self.cb_shift.set(match.group(1))
            self.ent_worktime.set(match.group(2).strip())
        else:
            # Fallback for old format (just time)
            self.cb_shift.set("주간")
            self.ent_worktime.set(str(val).strip())

    def get_ot(self): return self.ent_ot.get()
    def set_ot(self, val):
        self.ent_ot.delete(0, tk.END)
        self.ent_ot.insert(0, val)

    def bind_name(self, seq, func): self.cb_name.bind(seq, func)
    def bind_time(self, seq, func): 
        self.ent_worktime.bind(seq, func)
        # Shift selection should also trigger any auto-save bindings
        self.cb_shift.bind('<<ComboboxSelected>>', func, add='+')
        if 'FocusOut' in seq or 'Return' in seq:
            # Also trigger on selection from dropdown
            self.ent_worktime.bind('<<ComboboxSelected>>', func, add='+')
            
    def bind_ot(self, seq, func): self.ent_ot.bind(seq, func)

    def update_time_list(self, new_list):
        """Refresh the combobox values with a new list"""
        if hasattr(self, 'ent_worktime'):
            self.ent_worktime['values'] = new_list

class VehicleInspectionWidget(ttk.Frame):
    """
    Dedicated widget for vehicle inspection records with scrollable content
    """
    def __init__(self, parent, theme_bg='#f0f0f0', vehicle_list=None, **kwargs):
        super().__init__(parent, padding=0)
        
        # Create Canvas and Scrollbar for internal scrolling
        self.canvas = tk.Canvas(self, highlightthickness=0, bg=theme_bg)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        # Scroll binding is handled globally in MaterialManager.__init__

        def _on_canvas_configure(event):
            # Update width of inner frame to match canvas
            self.canvas.itemconfig(self.canvas_window, width=event.width)
        self.canvas.bind("<Configure>", _on_canvas_configure)

        # 1. Inspection Items (Checkboxes - Wrap in a grid for better horizontal usage)
        chk_frame = ttk.LabelFrame(self.scrollable_frame, text="점검 항목")
        chk_frame.pack(fill='x', padx=5, pady=2)
        
        self.vars = {}
        items = [("주유", "Fuel"), ("청결", "Clean"), ("엔진오일", "Oil"), ("타이어", "Tire"), ("전조등", "Light")]
        
        for i, (label, key) in enumerate(items):
            var = tk.BooleanVar()
            cb = ttk.Checkbutton(chk_frame, text=label, variable=var)
            # Use grid with multiple columns to handle narrow widths better
            cb.grid(row=i // 3, column=i % 3, padx=5, pady=2, sticky='w')
            self.vars[key] = var
            
        # 2. Input Fields
        input_frame = ttk.Frame(self.scrollable_frame)
        input_frame.pack(fill='x', padx=5, pady=2)
        
        ttk.Label(input_frame, text="차량정보:").grid(row=0, column=0, padx=2, pady=2, sticky='e')
        self.cb_vehicle_info = ttk.Combobox(input_frame, width=15)
        self.cb_vehicle_info.grid(row=0, column=1, padx=2, pady=2, sticky='w')
        if vehicle_list is not None:
            self.update_vehicle_list(vehicle_list)
            
        # Focus transitions
        def move_to_mileage(e): self.ent_mileage.focus_set()
        self.cb_vehicle_info.bind('<<ComboboxSelected>>', move_to_mileage)
        self.cb_vehicle_info.bind('<Return>', move_to_mileage)

        ttk.Label(input_frame, text="주행거리 (km):").grid(row=1, column=0, padx=2, pady=2, sticky='e')
        self.ent_mileage = ttk.Entry(input_frame, width=15)
        self.ent_mileage.grid(row=1, column=1, padx=2, pady=2, sticky='w')
        
        def on_mileage_return(e):
            self.format_mileage()
            self.ent_remarks.focus_set()
            
        self.ent_mileage.bind('<Return>', on_mileage_return)
        self.ent_mileage.bind('<FocusOut>', lambda e: self.format_mileage())

        ttk.Label(input_frame, text="비고:").grid(row=2, column=0, padx=2, pady=2, sticky='e')
        self.ent_remarks = ttk.Entry(input_frame)
        self.ent_remarks.grid(row=2, column=1, padx=2, pady=2, sticky='ew')
        input_frame.grid_columnconfigure(1, weight=1)

    def update_vehicle_list(self, new_list):
        """Update the combobox values and re-register autocomplete"""
        self.cb_vehicle_info['values'] = new_list
        # Access register_autocomplete from global scope
        try:
            import sys
            main_mod = sys.modules.get('__main__')
            if main_mod and hasattr(main_mod, 'register_autocomplete'):
                main_mod.register_autocomplete(self.cb_vehicle_info, new_list)
            elif 'register_autocomplete' in globals():
                globals()['register_autocomplete'](self.cb_vehicle_info, new_list)
        except:
            pass

    def format_mileage(self, event=None):
        """Autoformat mileage with commas"""
        try:
            val = self.ent_mileage.get().strip().replace(',', '')
            if not val: return
            # Handle possible float if someone type 123.4
            if '.' in val:
                num = float(val)
                formatted = f"{num:,.1f}"
                if formatted.endswith('.0'): formatted = formatted[:-2]
            else:
                num = int(val)
                formatted = f"{num:,}"
            
            self.ent_mileage.delete(0, tk.END)
            self.ent_mileage.insert(0, formatted)
        except:
            pass

    def get_data(self):
        data = {key: var.get() for key, var in self.vars.items()}
        data['vehicle_info'] = self.cb_vehicle_info.get()
        data['mileage'] = self.ent_mileage.get()
        data['remarks'] = self.ent_remarks.get()
        return data
        
    def set_data(self, data):
        if not data: return
        for key, var in self.vars.items():
            if key in data: var.set(data[key])
        if 'vehicle_info' in data:
            self.cb_vehicle_info.set(data['vehicle_info'])
        if 'mileage' in data:
            self.ent_mileage.delete(0, tk.END)
            self.ent_mileage.insert(0, data['mileage'])
        if 'remarks' in data:
            self.ent_remarks.delete(0, tk.END)
            self.ent_remarks.insert(0, data['remarks'])

def enable_column_resize(frame, num_cols, header_row=0, edge_px=6):
    """
    frame 내부 grid 테이블의 헤더 셀(header_row) 오른쪽 경계를 드래그해
    해당 컬럼과 오른쪽 인접 컬럼의 너비를 조절하는 기능을 추가합니다.

    - frame  : ttk.Frame, grid 레이아웃으로 컬럼이 구성된 프레임
    - num_cols : 컬럼 수
    - header_row : 헤더가 위치한 row 번호 (기본 0)
    - edge_px : 경계 감지 영역 (픽셀)
    """
    _drag = {'col': None, 'start_x': 0, 'start_w': 0, 'next_w': 0}

    def get_col_width(col):
        """현재 컬럼의 실제 픽셀 너비를 반환"""
        try:
            info = frame.grid_columnconfigure(col)
            minsz = info.get('minsize', 0)
            if minsz and minsz > 0:
                return minsz
            for widget in frame.grid_slaves(row=header_row, column=col):
                w = widget.winfo_width()
                if w > 1:
                    return w
            return 80
        except:
            return 80

    def on_motion(event):
        """드래그 중이 아닐 때만 커서 모양 변경"""
        if _drag['col'] is not None:
            return  # 이미 드래그 중이면 무시
        widget = event.widget
        x = event.x
        w = widget.winfo_width()
        if w - edge_px <= x <= w:
            widget.configure(cursor='sb_h_double_arrow')
        else:
            widget.configure(cursor='')

    def on_leave(event):
        if _drag['col'] is None:
            event.widget.configure(cursor='')

    def on_press(event):
        widget = event.widget
        x = event.x
        w = widget.winfo_width()
        if w - edge_px <= x <= w:
            info = widget.grid_info()
            col = info.get('column', -1)
            if col < 0 or col >= num_cols - 1:
                return
            _drag['col'] = col
            _drag['widget'] = widget
            _drag['start_x'] = event.x_root
            _drag['start_w'] = get_col_width(col)
            _drag['next_w'] = get_col_width(col + 1)
            # grab_set 제거 - 스크롤 방해 방지
        else:
            _drag['col'] = None

    def on_drag(event):
        if _drag['col'] is None:
            return
        col = _drag['col']
        dx = event.x_root - _drag['start_x']
        new_w = max(30, _drag['start_w'] + dx)
        new_next = max(30, _drag['next_w'] - dx)
        frame.grid_columnconfigure(col, minsize=new_w, weight=0)
        frame.grid_columnconfigure(col + 1, minsize=new_next, weight=0)

    def on_release(event):
        _drag['col'] = None
        # 커서 원상복구
        try:
            w = event.widget
            w.configure(cursor='')
        except:
            pass

    # 헤더 row(0) 의 모든 위젯에 바인딩
    def bind_headers():
        for col in range(num_cols):
            for widget in frame.grid_slaves(row=header_row, column=col):
                widget.bind('<Motion>', on_motion, add='+')
                widget.bind('<Leave>', on_leave, add='+')
                widget.bind('<ButtonPress-1>', on_press, add='+')
                widget.bind('<B1-Motion>', on_drag, add='+')
                widget.bind('<ButtonRelease-1>', on_release, add='+')

    # 위젯이 아직 렌더되기 전일 수 있으므로 idle 후 바인딩
    frame.after_idle(bind_headers)


class LaborCostDetailWidget(ttk.Frame):
    """
    Detailed labor cost calculation widget with three sections: 
    1) Regular Work (정시근무), 2) Special Work (특별근무), and 3) Base Salary Reference (기준급여)
    """
    def __init__(self, parent, on_change_callback=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.on_change_callback = on_change_callback
        
        # Rankings for Table 1
        self.ranks = ["이사", "부장", "차장", "과장", "대리", "계장", "주임", "기사"]
        # Shift types for Table 2
        self.special_types = ["연장근무", "야간근무", "휴일근무"]
        
        # Base Salaries for Reference (Table 3)
        self.base_salaries = {
            "이사": 55250000,
            "부장": 55250000,
            "차장": 47670000,
            "과장": 41170000,
            "대리": 37920000,
            "계장": 34670000,
            "주임": 31420000,
            "기사": 29250000
        }
        
        self.entries = {} # Key -> Rank/Type -> Column -> Entry
        self.totals = {}  # Key -> Rank/Type -> Label
        
        self._create_widgets()

    def _create_widgets(self):
        style = ttk.Style()
        style.configure("LaborHeader.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#e0e0e0', relief='solid')
        style.configure("LaborTotal.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#fff9c4', relief='solid')
        style.configure("RefTitle.TLabel", font=('Malgun Gothic', 10, 'bold'), foreground='red', background='#fce4ec', relief='solid')
        
        # Main Container
        main_container = ttk.Frame(self)
        main_container.pack(fill='both', expand=True)
        
        # Left Side (Calculation)
        calc_frame = ttk.Frame(main_container)
        calc_frame.pack(side='left', fill='both', expand=True)
        
        # Right Side (Reference Table)
        ref_frame = ttk.Frame(main_container, padding=(20, 40, 0, 0))
        ref_frame.pack(side='right', fill='y')

        # --- Section 1: Regular Work (Left) ---
        ttk.Label(calc_frame, text="1) 정시근무 (240일/년)", font=('Malgun Gothic', 11, 'bold')).pack(anchor='w', pady=(10, 5))
        
        table1_frame = ttk.Frame(calc_frame)
        table1_frame.pack(fill='x')
        
        headers1 = ["구분", "직급", "투입인원(명)", "투입일수/인(일)", "단가/일", "사전원가가액"]
        for j, h in enumerate(headers1):
            lbl = ttk.Label(table1_frame, text=h, style="LaborHeader.TLabel", padding=5, anchor='center')
            lbl.grid(row=0, column=j, sticky='nsew')
            table1_frame.grid_columnconfigure(j, weight=1)
        enable_column_resize(table1_frame, len(headers1))

        merge_lbl = ttk.Label(table1_frame, text="정시근무\n(240일/년)", relief='solid', anchor='center', padding=10)
        merge_lbl.grid(row=1, column=0, rowspan=len(self.ranks), sticky='nsew')

        for i, rank in enumerate(self.ranks):
            row = i + 1
            ttk.Label(table1_frame, text=rank, relief='solid', anchor='center', padding=5).grid(row=row, column=1, sticky='nsew')
            
            self.entries[rank] = {}
            ent_personnel = ttk.Entry(table1_frame, width=10, justify='center')
            ent_personnel.grid(row=row, column=2, sticky='nsew')
            ent_personnel.bind("<KeyRelease>", lambda e, r=rank: self._on_input_change(r))
            self.entries[rank]['personnel'] = ent_personnel
            
            ent_days = ttk.Entry(table1_frame, width=10, justify='center')
            ent_days.grid(row=row, column=3, sticky='nsew')
            ent_days.bind("<KeyRelease>", lambda e, r=rank: self._on_input_change(r))
            self.entries[rank]['period'] = ent_days
            
            ent_price = ttk.Entry(table1_frame, width=15, justify='right')
            ent_price.grid(row=row, column=4, sticky='nsew')
            ent_price.bind("<KeyRelease>", lambda e, r=rank: self._on_input_change(r))
            self.entries[rank]['unit_price'] = ent_price
            
            # [NEW] Default value from base salary / 240
            daily_rate = round(self.base_salaries.get(rank, 0) / 240)
            if daily_rate > 0:
                # Add a marker check to avoid overwriting user edits if necessary, 
                # but initially we can just populate it.
                ent_price.insert(0, f"{daily_rate:,.0f}")
            
            lbl_subtotal = ttk.Label(table1_frame, text="0", relief='solid', anchor='e', padding=5)
            lbl_subtotal.grid(row=row, column=5, sticky='nsew')
            self.totals[rank] = lbl_subtotal

        # Table 1 Subtotal Row
        row_t1_sum = len(self.ranks) + 1
        ttk.Label(table1_frame, text="소계", style="LaborHeader.TLabel", anchor='center', padding=5).grid(row=row_t1_sum, column=0, columnspan=2, sticky='nsew')
        self.lbl_t1_personnel_sum = ttk.Label(table1_frame, text="0 명", style="LaborHeader.TLabel", anchor='center')
        self.lbl_t1_personnel_sum.grid(row=row_t1_sum, column=2, sticky='nsew')
        self.lbl_t1_days_sum = ttk.Label(table1_frame, text="0 일", style="LaborHeader.TLabel", anchor='center')
        self.lbl_t1_days_sum.grid(row=row_t1_sum, column=3, sticky='nsew')
        ttk.Label(table1_frame, text="", style="LaborHeader.TLabel").grid(row=row_t1_sum, column=4, sticky='nsew')
        self.lbl_t1_cost_sum = ttk.Label(table1_frame, text="0", style="LaborHeader.TLabel", anchor='e', padding=5)
        self.lbl_t1_cost_sum.grid(row=row_t1_sum, column=5, sticky='nsew')

        # --- Section 2: Special Work (Left) ---
        ttk.Label(calc_frame, text="2) 특별근무", font=('Malgun Gothic', 11, 'bold')).pack(anchor='w', pady=(20, 5))
        
        table2_frame = ttk.Frame(calc_frame)
        table2_frame.pack(fill='x')
        
        headers2 = ["구분", "형태", "투입인원(명)", "투입시간/인(시간)", "단가", "사전원가가액"]
        for j, h in enumerate(headers2):
            lbl = ttk.Label(table2_frame, text=h, style="LaborHeader.TLabel", padding=5, anchor='center')
            lbl.grid(row=0, column=j, sticky='nsew')
            table2_frame.grid_columnconfigure(j, weight=1)
        enable_column_resize(table2_frame, len(headers2))

        merge_lbl2 = ttk.Label(table2_frame, text="특별근무", relief='solid', anchor='center', padding=10)
        merge_lbl2.grid(row=1, column=0, rowspan=len(self.special_types), sticky='nsew')

        for i, stype in enumerate(self.special_types):
            row = i + 1
            ttk.Label(table2_frame, text=stype, relief='solid', anchor='center', padding=5).grid(row=row, column=1, sticky='nsew')
            
            self.entries[stype] = {}
            ent_personnel = ttk.Entry(table2_frame, width=10, justify='center')
            ent_personnel.grid(row=row, column=2, sticky='nsew')
            ent_personnel.bind("<KeyRelease>", lambda e, s=stype: self._on_input_change(s))
            self.entries[stype]['personnel'] = ent_personnel
            
            ent_hours = ttk.Entry(table2_frame, width=10, justify='center')
            ent_hours.grid(row=row, column=3, sticky='nsew')
            ent_hours.bind("<KeyRelease>", lambda e, s=stype: self._on_input_change(s))
            self.entries[stype]['period'] = ent_hours
            
            ent_price = ttk.Entry(table2_frame, width=15, justify='right')
            ent_price.grid(row=row, column=4, sticky='nsew')
            ent_price.bind("<KeyRelease>", lambda e, s=stype: self._on_input_change(s))
            self.entries[stype]['unit_price'] = ent_price
            
            # [NEW] Default OT Rates
            ot_rates = {"연장근무": 4000, "야간근무": 5000, "휴일근무": 7500}
            if stype in ot_rates:
                ent_price.insert(0, f"{ot_rates[stype]:,.0f}")
            
            lbl_subtotal = ttk.Label(table2_frame, text="0", relief='solid', anchor='e', padding=5)
            lbl_subtotal.grid(row=row, column=5, sticky='nsew')
            self.totals[stype] = lbl_subtotal

        # Table 2 Subtotal Row
        row_t2_sum = len(self.special_types) + 1
        ttk.Label(table2_frame, text="소계", style="LaborHeader.TLabel", anchor='center', padding=5).grid(row=row_t2_sum, column=0, columnspan=2, sticky='nsew')
        self.lbl_t2_personnel_sum = ttk.Label(table2_frame, text="0 명", style="LaborHeader.TLabel", anchor='center')
        self.lbl_t2_personnel_sum.grid(row=row_t2_sum, column=2, sticky='nsew')
        self.lbl_t2_hours_sum = ttk.Label(table2_frame, text="0 시간", style="LaborHeader.TLabel", anchor='center')
        self.lbl_t2_hours_sum.grid(row=row_t2_sum, column=3, sticky='nsew')
        ttk.Label(table2_frame, text="", style="LaborHeader.TLabel").grid(row=row_t2_sum, column=4, sticky='nsew')
        self.lbl_t2_cost_sum = ttk.Label(table2_frame, text="0", style="LaborHeader.TLabel", anchor='e', padding=5)
        self.lbl_t2_cost_sum.grid(row=row_t2_sum, column=5, sticky='nsew')

        # --- Section 3: Base Salary Reference (Right) ---
        ttk.Label(ref_frame, text="고정금액(변경불가)", style="RefTitle.TLabel", padding=5).pack(fill='x')
        
        ref_table = ttk.Frame(ref_frame)
        ref_table.pack(fill='x')
        
        ttk.Label(ref_table, text="직급", style="LaborHeader.TLabel", padding=5, width=10, anchor='center').grid(row=0, column=0, sticky='nsew')
        ttk.Label(ref_table, text="기준급여", style="LaborHeader.TLabel", padding=5, width=15, anchor='center').grid(row=0, column=1, sticky='nsew')
        
        # Display ranks (skipping 이사 as it's not in the reference list image, but we have the data)
        display_ranks = ["부장", "차장", "과장", "대리", "계장", "주임", "기사"]
        for i, rank in enumerate(display_ranks):
            row = i + 1
            ttk.Label(ref_table, text=rank, relief='solid', padding=5, anchor='center').grid(row=row, column=0, sticky='nsew')
            salary = self.base_salaries.get(rank, 0)
            ttk.Label(ref_table, text=f"{salary:,.0f}", relief='solid', padding=5, anchor='e').grid(row=row, column=1, sticky='nsew')

        ttk.Button(ref_frame, text="기준급여 일괄 적용", command=self.apply_base_salaries).pack(pady=10, fill='x')

        # --- Grand Total ---
        total_frame = ttk.Frame(calc_frame)
        total_frame.pack(fill='x', pady=(10, 0))
        ttk.Label(total_frame, text="인건비 합계", style="LaborTotal.TLabel", anchor='center', padding=10).pack(side='left', fill='x', expand=True)
        self.lbl_grand_total = ttk.Label(total_frame, text="0", style="LaborTotal.TLabel", anchor='e', font=('Malgun Gothic', 12, 'bold'), padding=10)
        self.lbl_grand_total.pack(side='right', fill='x', expand=True)

    def apply_base_salaries(self):
        """Reset Daily Unit Price to the standard (Base Salary / 240)"""
        for rank, salary in self.base_salaries.items():
            if rank in self.entries:
                daily_rate = round(salary / 240)
                self.entries[rank]['unit_price'].delete(0, tk.END)
                self.entries[rank]['unit_price'].insert(0, f"{daily_rate:,.0f}")
                self._on_input_change(rank)
        messagebox.showinfo("적용 완료", "기준급여에 따른 일일 단가가 적용되었습니다.")

    def _to_f(self, val):
        try:
            return float(str(val).replace(',', '') or 0)
        except:
            return 0.0

    def _on_input_change(self, key):
        # Calculate row total
        personnel = self._to_f(self.entries[key]['personnel'].get())
        period = self._to_f(self.entries[key]['period'].get())
        price = self._to_f(self.entries[key]['unit_price'].get())
        
        row_total = personnel * period * price
        self.totals[key].config(text=f"{row_total:,.0f}")
        
        self.calculate_all()

    def calculate_all(self):
        # Table 1 Totals
        t1_personnel = 0
        t1_days = 0
        t1_cost = 0
        for rank in self.ranks:
            p = self._to_f(self.entries[rank]['personnel'].get())
            d = self._to_f(self.entries[rank]['period'].get())
            c = self._to_f(self.totals[rank].cget('text'))
            t1_personnel += p
            t1_days += d
            t1_cost += c
        
        self.lbl_t1_personnel_sum.config(text=f"{t1_personnel:g} 명")
        self.lbl_t1_days_sum.config(text=f"{t1_days:g} 일")
        self.lbl_t1_cost_sum.config(text=f"{t1_cost:,.0f}")

        # Table 2 Totals
        t2_personnel = 0
        t2_hours = 0
        t2_cost = 0
        for stype in self.special_types:
            p = self._to_f(self.entries[stype]['personnel'].get())
            h = self._to_f(self.entries[stype]['period'].get())
            c = self._to_f(self.totals[stype].cget('text'))
            t2_personnel += p
            t2_hours += h
            t2_cost += c
            
        self.lbl_t2_personnel_sum.config(text=f"{t2_personnel:g} 명")
        self.lbl_t2_hours_sum.config(text=f"{t2_hours:g} 시간")
        self.lbl_t2_cost_sum.config(text=f"{t2_cost:,.0f}")

        grand_total = t1_cost + t2_cost
        self.lbl_grand_total.config(text=f"₩ {grand_total:,.0f}")
        
        if self.on_change_callback:
            self.on_change_callback(grand_total)

    def get_data(self):
        """Export all entry values as a dictionary"""
        data = {}
        for key, widgets in self.entries.items():
            data[key] = {
                'personnel': widgets['personnel'].get(),
                'period': widgets['period'].get(),
                'unit_price': widgets['unit_price'].get()
            }
        return data

    def set_data(self, data):
        """Populate entries from a dictionary"""
        if not data or not isinstance(data, dict):
            self.reset()
            return
            
        for key, values in data.items():
            if key in self.entries:
                self.entries[key]['personnel'].delete(0, tk.END); self.entries[key]['personnel'].insert(0, values.get('personnel', ''))
                self.entries[key]['period'].delete(0, tk.END); self.entries[key]['period'].insert(0, values.get('period', ''))
                self.entries[key]['unit_price'].delete(0, tk.END); self.entries[key]['unit_price'].insert(0, values.get('unit_price', ''))
        
        # Trigger all row calculations
        for key in list(self.ranks) + list(self.special_types):
            self._on_input_change(key)

    def reset(self):
        """Clear all entries"""
        for key, widgets in self.entries.items():
            widgets['personnel'].delete(0, tk.END)
            widgets['period'].delete(0, tk.END)
            widgets['unit_price'].delete(0, tk.END)
        self.calculate_all()

class MaterialCostDetailWidget(ttk.Frame):
    """
    Detailed material cost calculation widget.
    Columns: Item (품목), Spec (사양), Quantity (수량), Unit (규격), Price (단가), Amount (사전원가가액)
    """
    def __init__(self, parent, on_change_callback=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.on_change_callback = on_change_callback
        
        # Default Items from Image
        self.default_items = [
            ("PT 약품", "세척액", "CAN", 1500),
            ("PT 약품", "침투액", "CAN", 2300),
            ("PT 약품", "현상액", "CAN", 2000),
            ("MT 약품", "자분페인트", "CAN", 2350),
            ("MT 약품", "흑색자분", "CAN", 1800),
            ("방사선투과검사 필름", "MX125", "매", 990),
            ("글리세린", "20L", "통", 100000),
            ("필름 현상액", "3L", "통", 16500),
            ("필름 정착액", "3L", "통", 16500),
            ("수적방지액", "200mL", "통", 2500)
        ]
        
        self.entries = [] # List of dicts for each row: {'item': lbl, 'spec': lbl, 'qty': ent, 'unit': lbl, 'price': ent, 'amount': lbl}
        self._create_widgets()

    def _create_widgets(self):
        style = ttk.Style()
        style.configure("MatHeader.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#e0e0e0', relief='solid')
        style.configure("MatTotal.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#ffff00', relief='solid') # Yellow as in image
        style.configure("MatRef.TLabel", font=('Malgun Gothic', 9), background='#f5f5f5', relief='solid')

        # Main Container
        main_container = ttk.Frame(self)
        main_container.pack(fill='both', expand=True)

        # Left Side (Calc)
        calc_frame = ttk.Frame(main_container)
        calc_frame.pack(side='left', fill='both', expand=True)

        # Right Side (Ref)
        ref_frame = ttk.Frame(main_container, padding=(20, 40, 0, 0))
        ref_frame.pack(side='right', fill='y')

        # --- Section 1: Tables (Left) ---
        ttk.Label(calc_frame, text="2) 재료비", font=('Malgun Gothic', 11, 'bold')).pack(anchor='w', pady=(10, 5))
        
        table_frame = ttk.Frame(calc_frame)
        table_frame.pack(fill='x')
        
        headers = ["품목", "사양", "수량", "규격", "단가", "사전원가가액"]
        widths = [20, 15, 10, 8, 15, 20]
        for j, (h, w) in enumerate(zip(headers, widths)):
            lbl = ttk.Label(table_frame, text=h, style="MatHeader.TLabel", padding=5, anchor='center', width=w)
            lbl.grid(row=0, column=j, sticky='nsew')
            table_frame.grid_columnconfigure(j, weight=1 if j in [0, 5] else 0)
        enable_column_resize(table_frame, len(headers))

        for i, (item, spec, unit, price) in enumerate(self.default_items):
            row = i + 1
            # Item
            ttk.Label(table_frame, text=item, relief='solid', padding=5, anchor='center').grid(row=row, column=0, sticky='nsew')
            # Spec
            ttk.Label(table_frame, text=spec, relief='solid', padding=5, anchor='center').grid(row=row, column=1, sticky='nsew')
            
            row_widgets = {}
            # Quantity
            ent_qty = ttk.Entry(table_frame, width=10, justify='center')
            ent_qty.grid(row=row, column=2, sticky='nsew')
            ent_qty.bind("<KeyRelease>", lambda e, idx=i: self._on_input_change(idx))
            row_widgets['qty'] = ent_qty
            
            # Unit
            ttk.Label(table_frame, text=unit, relief='solid', padding=5, anchor='center').grid(row=row, column=3, sticky='nsew')
            
            # Unit Price
            ent_price = ttk.Entry(table_frame, width=15, justify='right')
            ent_price.grid(row=row, column=4, sticky='nsew')
            ent_price.insert(0, f"{price:,.0f}")
            ent_price.bind("<KeyRelease>", lambda e, idx=i: self._on_input_change(idx))
            row_widgets['price'] = ent_price
            
            # Amount
            lbl_amount = ttk.Label(table_frame, text="0", relief='solid', anchor='e', padding=5)
            lbl_amount.grid(row=row, column=5, sticky='nsew')
            row_widgets['amount'] = lbl_amount
            
            self.entries.append(row_widgets)

        # Footer 1: Total
        row_footer = len(self.default_items) + 1
        ttk.Label(table_frame, text="재료비 합계", style="MatTotal.TLabel", anchor='center', padding=5).grid(row=row_footer, column=0, columnspan=5, sticky='nsew')
        self.lbl_mat_total = ttk.Label(table_frame, text="₩ 0", style="MatTotal.TLabel", anchor='e', padding=5)
        self.lbl_mat_total.grid(row=row_footer, column=5, sticky='nsew')

        # Footer 2: VAT
        row_vat = len(self.default_items) + 2
        ttk.Label(table_frame, text="부가세", relief='solid', anchor='center', padding=5).grid(row=row_vat, column=4, sticky='nsew')
        self.lbl_mat_vat = ttk.Label(table_frame, text="0", relief='solid', anchor='e', padding=5)
        self.lbl_mat_vat.grid(row=row_vat, column=5, sticky='nsew')

        # --- Section 2: Ref Info (Right) ---
        ref_table = ttk.Frame(ref_frame)
        ref_table.pack(fill='x')

        # Headers for reference
        ref_data = [
            ("지에스켐", "30M"),
            ("지에스켐", ""),
            ("지에스켐", ""),
            ("지에스켐", "32M"),
            ("지에스켐", ""),
            ("한스", "3 1/3 x 12\" (1850 매)"),
            ("한스", "3 1/3 x 6\" (313 매)"),
            ("한스", "총 매수: 2,163"),
            ("경도", "500매 기준 / 1회 4통 / 5회 교환"),
            ("나우", "500매 기준 / 1회 2동 / 5회 교환")
        ]

        for i, (title, val) in enumerate(ref_data):
            row = i
            ttk.Label(ref_table, text=title, width=10, anchor='center', style="MatRef.TLabel", padding=3).grid(row=row, column=0, sticky='nsew')
            ttk.Label(ref_table, text=val, width=25, anchor='w', style="MatRef.TLabel", padding=3).grid(row=row, column=1, sticky='nsew')

    def _on_input_change(self, idx):
        widgets = self.entries[idx]
        qty = self._to_f(widgets['qty'].get())
        price = self._to_f(widgets['price'].get())
        
        amount = qty * price
        widgets['amount'].config(text=f"{amount:,.0f}")
        
        self.calculate_all()

    def calculate_all(self):
        total_mat = 0.0
        for widgets in self.entries:
            qty = self._to_f(widgets['qty'].get())
            price = self._to_f(widgets['price'].get())
            amount = qty * price
            widgets['amount'].config(text=f"{amount:,.0f}")
            total_mat += amount
            
        vat = total_mat * 0.1
        self.lbl_mat_total.config(text=f"₩ {total_mat:,.0f}")
        self.lbl_mat_vat.config(text=f"{vat:,.0f}")
        
        if self.on_change_callback:
            self.on_change_callback(total_mat)

    def _to_f(self, val):
        try:
            return float(str(val).replace(',', '') or 0)
        except:
            return 0.0

    def get_data(self):
        """Export entry values"""
        data = []
        for widgets in self.entries:
            data.append({
                'qty': widgets['qty'].get(),
                'price': widgets['price'].get()
            })
        return data

    def set_data(self, data):
        """Populate entries"""
        if not data or not isinstance(data, list):
            self.reset()
            return
            
        for i, val in enumerate(data):
            if i < len(self.entries):
                self.entries[i]['qty'].delete(0, tk.END); self.entries[i]['qty'].insert(0, val.get('qty', ''))
                self.entries[i]['price'].delete(0, tk.END); self.entries[i]['price'].insert(0, val.get('price', ''))
        
        self.calculate_all()

    def reset(self):
        """Clear quantities and restore default prices"""
        for i, widgets in enumerate(self.entries):
            widgets['qty'].delete(0, tk.END)
            # Restore default price
            default_price = self.default_items[i][3]
            widgets['price'].delete(0, tk.END); widgets['price'].insert(0, f"{default_price:,.0f}")
            
        self.calculate_all()

class ExpenseProfitDetailWidget(ttk.Frame):
    """
    Comprehensive expense and profit calculation widget.
    Sections: 1) Site Expenses, 2) Rental, 3) Outsource, 4) Insurance, 5) Depreciation, 6) Indirect Cost, 7) Profit
    """
    def __init__(self, parent, on_change_callback=None, get_labor_total_func=None, get_material_total_func=None, get_revenue_func=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.on_change_callback = on_change_callback
        self.get_labor_total = get_labor_total_func
        self.get_material_total = get_material_total_func
        self.get_revenue = get_revenue_func
        
        # Data structures for entries
        self.entries = {
            'site_expense': [], # list of dicts
            'rental': [],
            'outsource': [],
            'depreciation': []
        }
        
        self._create_widgets()

    def _create_widgets(self):
        style = ttk.Style()
        style.configure("ExpHeader.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#e0e0e0', relief='solid')
        style.configure("ExpTotal.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#ffff00', relief='solid')
        style.configure("Margin.TLabel", font=('Malgun Gothic', 10, 'bold'), background='#90ee90', relief='solid') # Light green for profit

        # --- Section 1: Site Expenses ---
        ttk.Label(self, text="3) 경비", font=('Malgun Gothic', 11, 'bold')).pack(anchor='w', pady=(10, 5))
        
        s1_frame = ttk.LabelFrame(self, text="(1) 현장 경비")
        s1_frame.pack(fill='x', pady=5)
        
        self.s1_table = ttk.Frame(s1_frame)
        self.s1_table.pack(fill='x')
        
        headers = ["구분", "내용", "인원수", "수량", "규격", "단가", "사전원가가액"]
        widths = [15, 20, 8, 8, 8, 15, 20]
        for j, (h, w) in enumerate(zip(headers, widths)):
            ttk.Label(self.s1_table, text=h, style="ExpHeader.TLabel", padding=5, anchor='center', width=w).grid(row=0, column=j, sticky='nsew')
            self.s1_table.grid_columnconfigure(j, weight=1 if j in [1, 6] else 0)
        enable_column_resize(self.s1_table, len(headers))

        # Default Site Expenses
        defaults_s1 = [
            ("차량유지비", "주유, 수리, 통행, 주차 등", "N/A", 1, "일", 150000 // 30),
            ("소모품비", "장갑,일회용 작업복외", "N/A", 1, "일", 15000 // 30),
            ("복리후생비", "생수, 음료 외 기타", "N/A", 1, "일", 50000 // 30),
            ("Se-175", "방사성동위원소 구매", "N/A", 1, "EA", 10000000 // 280)
        ]
        
        for i, (cat, cont, ppl, qty, unit, price) in enumerate(defaults_s1):
            self._add_row_s1(cat, cont, ppl, qty, unit, price)

        # --- Section 2: Rental Costs ---
        s2_frame = ttk.LabelFrame(self, text="(2) 장비/차량 임차료")
        s2_frame.pack(fill='x', pady=5)
        
        self.s2_table = ttk.Frame(s2_frame)
        self.s2_table.pack(fill='x')
        
        headers2 = ["구분", "사양", "수량", "사용기간", "기간단위", "단가/월,대", "사전원가가액"]
        for j, h in enumerate(headers2):
            ttk.Label(self.s2_table, text=h, style="ExpHeader.TLabel", padding=5, anchor='center', width=widths[j]).grid(row=0, column=j, sticky='nsew')
            self.s2_table.grid_columnconfigure(j, weight=1 if j in [1, 6] else 0)
        enable_column_resize(self.s2_table, len(headers2))
            
        # Add 3 empty rows by default
        for _ in range(3):
            self._add_row_s2()

        # --- Section 3: Outsource Costs ---
        s3_frame = ttk.LabelFrame(self, text="(3) 외주비/잡급")
        s3_frame.pack(fill='x', pady=5)
        
        self.s3_table = ttk.Frame(s3_frame)
        self.s3_table.pack(fill='x')
        
        headers3 = ["구분", "작업내용", "공수", "단가", "사전원가가액"]
        widths3 = [15, 30, 10, 15, 20]
        for j, h in enumerate(headers3):
            ttk.Label(self.s3_table, text=h, style="ExpHeader.TLabel", padding=5, anchor='center', width=widths3[j]).grid(row=0, column=j, sticky='nsew')
            self.s3_table.grid_columnconfigure(j, weight=1 if j in [1, 4] else 0)
        enable_column_resize(self.s3_table, len(headers3))
            
        self._add_row_s3("케이엔디이", "방사선투과검사", 0, 15000)
        for _ in range(2): self._add_row_s3()

        # --- Section 4: Social Insurance ---
        s4_frame = ttk.Frame(self)
        s4_frame.pack(fill='x', pady=5)
        ttk.Label(s4_frame, text="(4) 4대 보험료", font=('Malgun Gothic', 10, 'bold')).pack(side='left', padx=5)
        
        insurance_table = ttk.Frame(self)
        insurance_table.pack(fill='x')
        headers4 = ["구분", "산출 기준", "산출 인건비", "단가(요율)", "사전원가가액"]
        for j, h in enumerate(headers4):
            ttk.Label(insurance_table, text=h, style="ExpHeader.TLabel", padding=5, anchor='center', width=widths[j] if j < len(widths) else 20).grid(row=0, column=j, sticky='nsew')
            insurance_table.grid_columnconfigure(j, weight=1 if j in [1, 4] else 0)
        enable_column_resize(insurance_table, len(headers4))
            
        ttk.Label(insurance_table, text="4대 보험료", relief='solid', padding=5, anchor='center').grid(row=1, column=0, sticky='nsew')
        ttk.Label(insurance_table, text="산출인건비 X 요율(2024.7.1 기준)", relief='solid', padding=5, anchor='w').grid(row=1, column=1, sticky='nsew')
        self.lbl_insurance_base = ttk.Label(insurance_table, text="₩ 0", relief='solid', padding=5, anchor='e')
        self.lbl_insurance_base.grid(row=1, column=2, sticky='nsew')
        ttk.Label(insurance_table, text="10.6661%", relief='solid', padding=5, anchor='center').grid(row=1, column=3, sticky='nsew')
        self.lbl_insurance_amount = ttk.Label(insurance_table, text="0", relief='solid', padding=5, anchor='e')
        self.lbl_insurance_amount.grid(row=1, column=4, sticky='nsew')

        # --- Section 5: Depreciation ---
        s5_frame = ttk.LabelFrame(self, text="(5) 감기상각비 (Depreciation)")
        s5_frame.pack(fill='x', pady=5)
        
        self.s5_table = ttk.Frame(s5_frame)
        self.s5_table.pack(fill='x')
        
        headers5 = ["장비명", "사양", "내용년수", "수량", "사용일수", "감기비/일", "사전원가가액"]
        for j, h in enumerate(headers5):
            ttk.Label(self.s5_table, text=h, style="ExpHeader.TLabel", padding=5, anchor='center', width=widths[j]).grid(row=0, column=j, sticky='nsew')
            self.s5_table.grid_columnconfigure(j, weight=1 if j in [1, 6] else 0)
        enable_column_resize(self.s5_table, len(headers5))
            
        defaults_s5 = [
            ("PAUT 장비", "", 5, 1, 0, 44444),
            ("PAUT SCANNER (MANUAL)", "", 5, 1, 0, 5556),
            ("PAUT SCANNER (COBRA)", "", 5, 1, 0, 16667),
            ("YOKE", "", 5, 1, 0, 222),
            ("현상용 탑차(5년간 보험비 포함)", "현장별 차량기입시 탑차 구분 기입", 5, 1, 0, 16667),
            ("스타렉스(5년간 보험비 포함)", "현장별 차량기입시 스타렉스 구분 기입", 5, 1, 0, 16667)
        ]
        for item, spec, life, qty, days, rate in defaults_s5:
            self._add_row_s5(item, spec, life, qty, days, rate)

        # --- TOTALS SUMMARY ---
        summary_frame = ttk.Frame(self, padding=10)
        summary_frame.pack(fill='x', pady=10)
        
        # Row: Expense Total (1~5)
        ttk.Label(summary_frame, text="경비 합계 : (1)~(5) 합계", style="ExpTotal.TLabel", anchor='center', padding=8).grid(row=0, column=0, columnspan=4, sticky='nsew')
        self.lbl_exp_total = ttk.Label(summary_frame, text="₩ 0", style="ExpTotal.TLabel", anchor='e', padding=8)
        self.lbl_exp_total.grid(row=0, column=4, sticky='nsew')
        
        # Row: Expense VAT
        ttk.Label(summary_frame, text="경비 부가세 합계", relief='solid', anchor='center', padding=5).grid(row=1, column=3, sticky='nsew')
        self.lbl_exp_vat = ttk.Label(summary_frame, text="0", relief='solid', anchor='e', padding=5)
        self.lbl_exp_vat.grid(row=1, column=4, sticky='nsew')

        # Row: Total Direct Cost (Sales Cost)
        ttk.Label(summary_frame, text="매출원가 총계 : 1), 2), 3) 합계", font=('Malgun Gothic', 10, 'bold'), background='#00ffff', relief='solid', anchor='center', padding=8).grid(row=2, column=0, columnspan=4, sticky='nsew')
        self.lbl_sales_cost_total = ttk.Label(summary_frame, text="₩ 0", font=('Malgun Gothic', 10, 'bold'), background='#00ffff', relief='solid', anchor='e', padding=8)
        self.lbl_sales_cost_total.grid(row=2, column=4, sticky='nsew')

        # Row: Indirect Cost (판관비)
        ttk.Label(summary_frame, text="3. 간접비(판관비)", font=('Malgun Gothic', 10, 'bold'), anchor='w').grid(row=3, column=0, pady=(10, 0))
        
        indirect_table = ttk.Frame(summary_frame)
        indirect_table.grid(row=4, column=0, columnspan=5, sticky='ew')
        headers_ind = ["구분", "산출 기준", "산출직접비", "간접비율(%)", "사전 간접비 합계"]
        for j, h in enumerate(headers_ind):
             ttk.Label(indirect_table, text=h, style="ExpHeader.TLabel", padding=5, anchor='center', width=widths[j] if j < len(widths) else 25).grid(row=0, column=j, sticky='nsew')
             indirect_table.grid_columnconfigure(j, weight=1 if j in [1, 4] else 0)
        enable_column_resize(indirect_table, len(headers_ind))
        
        ttk.Label(indirect_table, text="간접비", relief='solid', padding=5, anchor='center').grid(row=1, column=0, sticky='nsew')
        ttk.Label(indirect_table, text="산출직접비 x 간접비율(2024년 기준)", relief='solid', padding=5, anchor='w').grid(row=1, column=1, sticky='nsew')
        self.lbl_indirect_base = ttk.Label(indirect_table, text="₩ 0", relief='solid', padding=5, anchor='e')
        self.lbl_indirect_base.grid(row=1, column=2, sticky='nsew')
        ttk.Label(indirect_table, text="14%", relief='solid', padding=5, anchor='center').grid(row=1, column=3, sticky='nsew')
        self.lbl_indirect_total = ttk.Label(indirect_table, text="₩ 0", font=('Malgun Gothic', 10, 'bold'), background='#00ffff', relief='solid', anchor='e', padding=5)
        self.lbl_indirect_total.grid(row=1, column=4, sticky='nsew')

        # Row: Total Cost (Direct + Indirect)
        ttk.Label(summary_frame, text="4. 총원가(매출원가+간접비)", style="ExpTotal.TLabel", anchor='center', padding=10).grid(row=5, column=0, columnspan=4, sticky='nsew', pady=(10, 0))
        self.lbl_grand_total_cost = ttk.Label(summary_frame, text="₩ 0", style="ExpTotal.TLabel", anchor='e', padding=10)
        self.lbl_grand_total_cost.grid(row=5, column=4, sticky='nsew', pady=(10, 0))

        # Row: Operating Profit Section
        ttk.Label(summary_frame, text="5. 영업이익, 영업이익률", font=('Malgun Gothic', 10, 'bold'), anchor='w').grid(row=6, column=0, pady=(10, 0))
        
        profit_table = ttk.Frame(summary_frame)
        profit_table.grid(row=7, column=0, columnspan=5, sticky='ew')
        headers_prof = ["구분", "매출(수입)", "총원가", "영업이익", "영업이익률", "기준"]
        widths_prof = [15, 20, 20, 20, 15, 10]
        for j, h in enumerate(headers_prof):
             ttk.Label(profit_table, text=h, style="ExpHeader.TLabel", padding=5, anchor='center', width=widths_prof[j]).grid(row=0, column=j, sticky='nsew')
             profit_table.grid_columnconfigure(j, weight=1 if j != 5 else 0)
        enable_column_resize(profit_table, len(headers_prof))

        # Row 1: Budget (사전)
        ttk.Label(profit_table, text="사전(예산)", relief='solid', padding=5, anchor='center').grid(row=1, column=0, sticky='nsew')
        self.lbl_prof_revenue = ttk.Label(profit_table, text="₩ 0", relief='solid', padding=5, anchor='e')
        self.lbl_prof_revenue.grid(row=1, column=1, sticky='nsew')
        self.lbl_prof_total_cost = ttk.Label(profit_table, text="₩ 0", relief='solid', padding=5, anchor='e')
        self.lbl_prof_total_cost.grid(row=1, column=2, sticky='nsew')
        self.lbl_prof_op_profit = ttk.Label(profit_table, text="₩ 0", style="Margin.TLabel", padding=5, anchor='center')
        self.lbl_prof_op_profit.grid(row=1, column=3, sticky='nsew')
        self.lbl_prof_margin = ttk.Label(profit_table, text="0.00%", relief='solid', padding=5, anchor='center')
        self.lbl_prof_margin.grid(row=1, column=4, sticky='nsew')
        ttk.Label(profit_table, text="부가세 별도", relief='solid', padding=5, anchor='center').grid(row=1, column=5, sticky='nsew')

    def _add_row_s1(self, cat="", cont="", ppl="", qty="", unit="", price=0):
        row = len(self.entries['site_expense']) + 1
        widgets = {}
        ent_cat = ttk.Entry(self.s1_table, width=15, justify='center'); ent_cat.insert(0, cat); ent_cat.grid(row=row, column=0, sticky='nsew')
        ent_cont = ttk.Entry(self.s1_table, width=20); ent_cont.insert(0, cont); ent_cont.grid(row=row, column=1, sticky='nsew')
        ent_ppl = ttk.Entry(self.s1_table, width=8, justify='center'); ent_ppl.insert(0, str(ppl)); ent_ppl.grid(row=row, column=2, sticky='nsew')
        ent_qty = ttk.Entry(self.s1_table, width=8, justify='center'); ent_qty.insert(0, str(qty)); ent_qty.grid(row=row, column=3, sticky='nsew')
        ent_unit = ttk.Entry(self.s1_table, width=8, justify='center'); ent_unit.insert(0, unit); ent_unit.grid(row=row, column=4, sticky='nsew')
        ent_price = ttk.Entry(self.s1_table, width=15, justify='right'); ent_price.insert(0, f"{price:,.0f}"); ent_price.grid(row=row, column=5, sticky='nsew')
        lbl_amt = ttk.Label(self.s1_table, text="0", relief='solid', anchor='e', padding=5); lbl_amt.grid(row=row, column=6, sticky='nsew')
        
        widgets = {'cat': ent_cat, 'cont': ent_cont, 'ppl': ent_ppl, 'qty': ent_qty, 'unit': ent_unit, 'price': ent_price, 'amount': lbl_amt}
        for w in [ent_ppl, ent_qty, ent_price]: w.bind("<KeyRelease>", lambda e: self.calculate_all())
        self.entries['site_expense'].append(widgets)

    def _add_row_s2(self, cat="", spec="", qty="", period="", unit="", price=0):
        row = len(self.entries['rental']) + 1
        widgets = {}
        e1 = ttk.Entry(self.s2_table, width=15, justify='center'); e1.insert(0, cat); e1.grid(row=row, column=0, sticky='nsew')
        e2 = ttk.Entry(self.s2_table, width=20); e2.insert(0, spec); e2.grid(row=row, column=1, sticky='nsew')
        e3 = ttk.Entry(self.s2_table, width=8, justify='center'); e3.insert(0, str(qty)); e3.grid(row=row, column=2, sticky='nsew')
        e4 = ttk.Entry(self.s2_table, width=8, justify='center'); e4.insert(0, str(period)); e4.grid(row=row, column=3, sticky='nsew')
        e5 = ttk.Entry(self.s2_table, width=8, justify='center'); e5.insert(0, unit); e5.grid(row=row, column=4, sticky='nsew')
        e6 = ttk.Entry(self.s2_table, width=15, justify='right'); e6.insert(0, f"{price:,.0f}"); e6.grid(row=row, column=5, sticky='nsew')
        lbl = ttk.Label(self.s2_table, text="0", relief='solid', anchor='e', padding=5); lbl.grid(row=row, column=6, sticky='nsew')
        
        widgets = {'cat': e1, 'spec': e2, 'qty': e3, 'period': e4, 'unit': e5, 'price': e6, 'amount': lbl}
        for w in [e3, e4, e6]: w.bind("<KeyRelease>", lambda e: self.calculate_all())
        self.entries['rental'].append(widgets)

    def _add_row_s3(self, cat="", work="", count=0, price=0):
        row = len(self.entries['outsource']) + 1
        widgets = {}
        e1 = ttk.Entry(self.s3_table, width=15, justify='center'); e1.insert(0, cat); e1.grid(row=row, column=0, sticky='nsew')
        e2 = ttk.Entry(self.s3_table, width=30); e2.insert(0, work); e2.grid(row=row, column=1, sticky='nsew')
        e3 = ttk.Entry(self.s3_table, width=10, justify='center'); e3.insert(0, str(count)); e3.grid(row=row, column=2, sticky='nsew')
        e4 = ttk.Entry(self.s3_table, width=15, justify='right'); e4.insert(0, f"{price:,.0f}"); e4.grid(row=row, column=3, sticky='nsew')
        lbl = ttk.Label(self.s3_table, text="0", relief='solid', anchor='e', padding=5); lbl.grid(row=row, column=4, sticky='nsew')
        
        widgets = {'cat': e1, 'work': e2, 'count': e3, 'price': e4, 'amount': lbl}
        for w in [e3, e4]: w.bind("<KeyRelease>", lambda e: self.calculate_all())
        self.entries['outsource'].append(widgets)

    def _add_row_s5(self, item="", spec="", life=5, qty=1, days=0, rate=0):
        row = len(self.entries['depreciation']) + 1
        widgets = {}
        e1 = ttk.Entry(self.s5_table, width=20); e1.insert(0, item); e1.grid(row=row, column=0, sticky='nsew')
        e2 = ttk.Entry(self.s5_table, width=15); e2.insert(0, spec); e2.grid(row=row, column=1, sticky='nsew')
        e3 = ttk.Entry(self.s5_table, width=8, justify='center'); e3.insert(0, str(life)); e3.grid(row=row, column=2, sticky='nsew')
        e4 = ttk.Entry(self.s5_table, width=8, justify='center'); e4.insert(0, str(qty)); e4.grid(row=row, column=3, sticky='nsew')
        e5 = ttk.Entry(self.s5_table, width=8, justify='center'); e5.insert(0, str(days)); e5.grid(row=row, column=4, sticky='nsew')
        e6 = ttk.Entry(self.s5_table, width=15, justify='right'); e6.insert(0, f"{rate:,.0f}"); e6.grid(row=row, column=5, sticky='nsew')
        lbl = ttk.Label(self.s5_table, text="0", relief='solid', anchor='e', padding=5); lbl.grid(row=row, column=6, sticky='nsew')
        
        widgets = {'item': e1, 'spec': e2, 'life': e3, 'qty': e4, 'days': e5, 'rate': e6, 'amount': lbl}
        for w in [e4, e5, e6]: w.bind("<KeyRelease>", lambda e: self.calculate_all())
        self.entries['depreciation'].append(widgets)

    def calculate_all(self, event=None):
        # 1. Site Expenses
        t1 = 0.0
        for w in self.entries['site_expense']:
            amt = self._to_f(w['qty'].get()) * self._to_f(w['price'].get())
            w['amount'].config(text=f"{amt:,.0f}")
            t1 += amt
            
        # 2. Rentals
        t2 = 0.0
        for w in self.entries['rental']:
            amt = self._to_f(w['qty'].get()) * self._to_f(w['period'].get()) * self._to_f(w['price'].get())
            w['amount'].config(text=f"{amt:,.0f}")
            t2 += amt
            
        # 3. Outsource
        t3 = 0.0
        for w in self.entries['outsource']:
            amt = self._to_f(w['count'].get()) * self._to_f(w['price'].get())
            w['amount'].config(text=f"{amt:,.0f}")
            t3 += amt
            
        # 4. Insurance
        labor_total = self.get_labor_total() if self.get_labor_total else 0.0
        t4 = labor_total * 0.106661
        self.lbl_insurance_base.config(text=f"₩ {labor_total:,.0f}")
        self.lbl_insurance_amount.config(text=f"{t4:,.0f}")
        
        # 5. Depreciation
        t5 = 0.0
        for w in self.entries['depreciation']:
            amt = self._to_f(w['qty'].get()) * self._to_f(w['days'].get()) * self._to_f(w['rate'].get())
            w['amount'].config(text=f"{amt:,.0f}")
            t5 += amt
            
        exp_total = t1 + t2 + t3 + t4 + t5
        exp_vat = (t1 + t2 + t3) * 0.1 # Example: VAT on direct expenses
        self.lbl_exp_total.config(text=f"₩ {exp_total:,.0f}")
        self.lbl_exp_vat.config(text=f"{exp_vat:,.0f}")
        
        # 6. Sales Cost (Labor + Material + Exp)
        mat_total = self.get_material_total() if self.get_material_total else 0.0
        direct_cost = labor_total + mat_total + exp_total
        self.lbl_sales_cost_total.config(text=f"₩ {direct_cost:,.0f}")
        
        # 7. Indirect Cost (14%)
        self.lbl_indirect_base.config(text=f"₩ {direct_cost:,.0f}")
        indirect_cost = direct_cost * 0.14
        self.lbl_indirect_total.config(text=f"₩ {indirect_cost:,.0f}")
        
        grand_total_cost = direct_cost + indirect_cost
        self.lbl_grand_total_cost.config(text=f"₩ {grand_total_cost:,.0f}")
        
        # 8. Profit
        revenue = self.get_revenue() if self.get_revenue else 0.0
        op_profit = revenue - grand_total_cost
        margin = (op_profit / revenue * 100) if revenue > 0 else 0.0
        
        self.lbl_prof_revenue.config(text=f"₩ {revenue:,.0f}")
        self.lbl_prof_total_cost.config(text=f"₩ {grand_total_cost:,.0f}")
        self.lbl_prof_op_profit.config(text=f"₩ {op_profit:,.0f}")
        self.lbl_prof_margin.config(text=f"{margin:.2f}%")
        
        # Update main form "Expense" and "Outsource" fields
        if self.on_change_callback:
            # We pass (Expense, Outsource, TotalProfit) or something?
            # Let's just update the main Expense field with exp_total (1~5)
            # and Outsource with t3.
            self.on_change_callback(exp_total, t3, op_profit)

    def _to_f(self, val):
        try:
            return float(str(val).replace(',', '') or 0)
        except:
            return 0.0

    def get_data(self):
        data = {
            'site_expense': [{k: v.get() if hasattr(v, 'get') else v.cget('text') for k, v in row.items()} for row in self.entries['site_expense']],
            'rental': [{k: v.get() if hasattr(v, 'get') else v.cget('text') for k, v in row.items()} for row in self.entries['rental']],
            'outsource': [{k: v.get() if hasattr(v, 'get') else v.cget('text') for k, v in row.items()} for row in self.entries['outsource']],
            'depreciation': [{k: v.get() if hasattr(v, 'get') else v.cget('text') for k, v in row.items()} for row in self.entries['depreciation']]
        }
        return data

    def set_data(self, data):
        if not data or not isinstance(data, dict):
            self.reset()
            return
            
        def fill(entry_list, data_list):
            for i, d in enumerate(data_list):
                if i < len(entry_list):
                    for k, v in d.items():
                        if k in entry_list[i] and hasattr(entry_list[i][k], 'delete'):
                            entry_list[i][k].delete(0, tk.END); entry_list[i][k].insert(0, str(v))
        
        fill(self.entries['site_expense'], data.get('site_expense', []))
        fill(self.entries['rental'], data.get('rental', []))
        fill(self.entries['outsource'], data.get('outsource', []))
        fill(self.entries['depreciation'], data.get('depreciation', []))
        
        self.calculate_all()

    def reset(self):
        def clear(entry_list):
            for row in entry_list:
                for k, v in row.items():
                    if hasattr(v, 'delete'): v.delete(0, tk.END)
        
        clear(self.entries['site_expense'])
        clear(self.entries['rental'])
        clear(self.entries['outsource'])
        clear(self.entries['depreciation'])
        self.calculate_all()

class ColumnSelectionDialog(tk.Toplevel):
    """Dialog to select columns for Excel export"""
    def __init__(self, parent, columns, title="엑셀 출력 컬럼 선택"):
        super().__init__(parent)
        self.title(title)
        self.geometry("500x700") # Larger for better visibility
        self.transient(parent)
        self.grab_set()
        
        self.result = None
        self.vars = {}
        
        # Header
        lbl = ttk.Label(self, text="출력할 컬럼을 선택하세요:", font=('Malgun Gothic', 11, 'bold'))
        lbl.pack(pady=10)
        
        # Scrollable area for checkboxes
        container = ttk.Frame(self)
        container.pack(fill='both', expand=True, padx=20)
        
        canvas = tk.Canvas(container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Unbind when closed to avoid errors
        def on_close():
            self.destroy()
        
        self.protocol("WM_DELETE_WINDOW", on_close)
            
        # Create checkboxes
        for col in columns:
            var = tk.BooleanVar(value=True)
            self.vars[col] = var
            # [NEW] Map 검사량 to 수량 for display
            display_map = {'검사량': '수량'}
            display_text = display_map.get(col, col)
            cb = ttk.Checkbutton(scrollable_frame, text=display_text, variable=var)
            cb.pack(anchor='w', pady=2)
            
        # Buttons area
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill='x', pady=15, padx=20)
        
        def select_all():
            for v in self.vars.values(): v.set(True)
        def deselect_all():
            for v in self.vars.values(): v.set(False)
            
        ttk.Button(btn_frame, text="전체 선택", command=select_all).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="전체 해제", command=deselect_all).pack(side='left', padx=5)
        
        # Bottom controls
        bottom_frame = ttk.Frame(self)
        bottom_frame.pack(fill='x', pady=10)
        
        def on_ok():
            canvas.unbind_all("<MouseWheel>") # Unbind on OK too
            self.result = [col for col, var in self.vars.items() if var.get()]
            self.destroy()
            
        def on_cancel():
            on_close()
            
        ttk.Button(bottom_frame, text="확인", command=on_ok).pack(side='right', padx=10)
        ttk.Button(bottom_frame, text="취소", command=on_close).pack(side='right', padx=5)
        
        # self.wait_window() removed from here to allow caller to set vars before blocking
        # self.wait_window() removed from here to allow caller to set vars before blocking

class MaterialManager:
    def __init__(self, root):
        self.root = root
        
        # High DPI awareness
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            try:
                ctypes.windll.user32.SetProcessDPIAware()
            except Exception:
                pass
                
        self.root.title("자재 및 소모품 관리 시스템 (Material Manager)")
        self.root.geometry("1600x900")
        try:
            self.root.state('zoomed') # Maximize on Windows
        except:
            pass
        
        # Configure overall style
        
        # Global Scroll Handler for all canvases in the app
        def _on_global_mousewheel(event):
            try:
                widget = event.widget.winfo_containing(event.x_root, event.y_root)
                if not widget: return
                
                # Do not interfere with native scrolling widgets
                if isinstance(widget, (tk.Text, ttk.Treeview, tk.Listbox)):
                    return
                    
                parent = widget
                while parent:
                    if isinstance(parent, tk.Canvas):
                        # Do not scroll the daily usage entry form canvas as it has no scrollbars
                        if hasattr(self, 'entry_canvas') and parent is self.entry_canvas:
                            return
                            
                        parent.yview_scroll(int(-1 * (event.delta / 120)), "units")
                        return
                    if hasattr(parent, 'master') and parent.master:
                        parent = parent.master
                    else:
                        break
            except Exception:
                pass
                
        self.root.bind_all("<MouseWheel>", _on_global_mousewheel)
        self.style = ttk.Style()
        try:
            self.style.theme_use('clam') # Use 'clam' theme for better grid line visibility
        except:
            pass
            
        self.style.configure(".", font=('Malgun Gothic', 12))
        self.style.configure("Treeview.Heading", font=('Malgun Gothic', 12, 'bold'))
        self.style.configure("Treeview", font=('Malgun Gothic', 12), rowheight=35) # Increased row height for "boxed" look
        
        # Detect system background color for tk widgets (Canvas, Text)
        self.theme_bg = self.root.cget('bg')
        if not self.theme_bg or self.theme_bg == 'SystemButtonFace':
             # Fallback/Modern check for Windows
             try:
                 self.theme_bg = self.style.lookup('TFrame', 'background')
             except:
                 self.theme_bg = '#f0f0f0' # Standard Windows gray
        
        # Determine base directory and bundle directory for portability
        if getattr(sys, 'frozen', False):
            # If running as an executable
            self.app_dir = os.path.dirname(sys.executable)
            self.bundle_dir = getattr(sys, '_MEIPASS', self.app_dir)
            # For exe, use user's Documents folder for config to ensure write permissions
            documents_dir = os.path.join(os.path.expanduser('~'), 'Documents', 'MaterialManager')
            if not os.path.exists(documents_dir):
                os.makedirs(documents_dir, exist_ok=True)
            self.config_path = os.path.join(documents_dir, 'Material_Manager_Config.json')
        else:
            # 스크립트 실행 모드: src/ 우선, 없으면 ../data/ 탐색
            self.app_dir = os.path.dirname(os.path.abspath(__file__))
            self.bundle_dir = self.app_dir
            self.config_path = os.path.join(self.app_dir, 'Material_Manager_Config.json')
            _db_name = 'Material_Inventory.xlsx'
            _data_dir = os.path.join(os.path.dirname(self.app_dir), 'data')
            _candidates = [
                os.path.join(self.app_dir, _db_name),   # src/Material_Inventory.xlsx
                os.path.join(_data_dir, _db_name),      # data/Material_Inventory.xlsx
            ]
            self.db_path = next((p for p in _candidates if os.path.exists(p)), _candidates[0])

        if getattr(sys, 'frozen', False):
            # exe 실행 모드: exe와 같은 폴더에 DB 저장 (어디서 실행해도 동일)
            self.db_path = os.path.join(self.app_dir, 'Material_Inventory.xlsx')

        # db_path 확정
        
        self.sites = [] # Initialize site list
        self.users = [
            "부장 주진철", "대리 우명광", "주임 김진환", "계장 장승대", "주임 김성렬", "부장 박광복", "과장 주영광"
        ] # Initialize worker/name list
        self.warehouses = [] # Initialize warehouse list
        self.equipments = [] # Initialize equipment list
        self.vehicles = [] # Initialize vehicle list
        self.worktimes = [] # Initialize worktimes list
        self.ot_times = [] # Initialize ot_times list
        
        # [NEW] Pre-initialize autocomplete/display lists to prevent AttributeError during load_data/refresh_filters
        self.materials_display_list = []
        self.co_code_list = []
        self.eq_code_list = []
        self.item_name_list = []
        self.class_list = []
        self.spec_list = []
        self.unit_list = []
        self.mfr_list = []
        self.origin_list = []
        self.sn_list = []
        self.model_list = []
        self.equipment_suggestions = []
        self.budget_sites = []  # [NEW] 공사실행예산서 전용 현장 목록 (budget_df 현장만)
        self.hidden_sites = []  # [NEW] 현장명 드롭다운에서 숨길 현장 목록 (config에 저장)
        
        self.load_data()
        
        # Centralized list of Carestream films for suggestions
        self.carestream_films = [
            "Carestream AA400-3⅓*12\"",
            "Carestream AA400-3⅓*17\"", 
            "Carestream AA400-4½*12\"",
            "Carestream AA400-10*12\"",
            "Carestream M100-3⅓*12\"",
            "Carestream M100-10*12\"",
            "Carestream M100-14*17\"",
            "Carestream MX125-3⅓*6\"",
            "Carestream MX125-3⅓*12\"",
            "Carestream MX125-4½*12\"",
            "Carestream MX125-10*12\"",
            "Carestream MX125-14*17\"",
            "Carestream T200-3⅓*12\"",
            "Carestream T200-3⅓*17\"",
            "Carestream T200-4½*12\"",
            "Carestream T200-10*12\"",
            "Carestream T200-14*17\""
        ]
        
        # Core draggable widget keys that should never be hidden
        self.CORE_DRAGGABLE_KEYS = [
            'form_box_geometry', 
            'ndt_usage_box_geometry', 
            'rtk_usage_box_geometry', 
            'save_btn_geometry', 
            'workers_box_geometry'
        ]
        
        # Registry for draggable items {config_key: widget_instance}
        self.draggable_items = {}
        self.memos = {} # key -> {'container': container, 'text_widget': text, 'title_entry': entry}
        self.checklists = {} # key -> {'container': container, 'title_entry': entry, 'item_frame': frame, 'items': []}
        self.vehicle_inspections = {} # key -> widget
        self.layout_locked = False
        self.daily_usage_sash_locked = False
        self._last_motion_time = 0 # For performance throttling
        self.is_ready = False  # Suppress saves until fully loaded

        # [CRITICAL FIX] Pre-load configuration synchronously to ensure UI starts in correct state
        self.preload_config_locks()

        self.create_widgets()
        self.update_registration_combos()
        
        # Enable keyboard navigation
        self.setup_keyboard_shortcuts()

    def preload_config_locks(self):
        """Pre-load critical lock states synchronously before UI creation"""
        try:
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.layout_locked = config.get('layout_locked', False)
                    self.daily_usage_sash_locked = config.get('daily_usage_sash_locked', False)
                    print(f"DEBUG: Pre-loaded locks - Layout: {self.layout_locked}, Sash: {self.daily_usage_sash_locked}")
        except Exception as e:
            print(f"DEBUG: Failed to pre-load locks: {e}")

    def _ensure_canvas_scroll_region(self):
        """Update canvas scroll region based on all children (grid and place)"""
        try:
            if hasattr(self, 'entry_canvas') and self.entry_canvas:
                self.entry_canvas.update_idletasks()
                # Get current bounding box of all items in the inner frame
                bbox = self.entry_canvas.bbox("all")
                if bbox:
                    # Add padding to ensure nothing is cut off. Width is at least 1100 for grid.
                    scroll_w = max(1100, bbox[2])
                    scroll_h = bbox[3] + 150
                    self.entry_canvas.configure(scrollregion=(0, 0, scroll_w, scroll_h))
                else:
                    # Fallback if no items
                    self.entry_canvas.configure(scrollregion=(0, 0, 2000, 2000))
        except:
            pass

    def _on_daily_usage_sash_changed(self, event=None):
        """Handle sash position change to save ratio"""
        try:
            # Skip saving if locked
            if hasattr(self, 'daily_usage_sash_locked') and self.daily_usage_sash_locked:
                return

            if hasattr(self, 'daily_usage_paned'):
                self.daily_usage_paned.update_idletasks()
                total_h = self.daily_usage_paned.winfo_height()
                if total_h > 0:
                    sash_pos = self.daily_usage_paned.sashpos(0)
                    ratio = sash_pos / total_h
                    
                    if not hasattr(self, 'tab_config'):
                        self.tab_config = {}
                    
                    self.tab_config['daily_usage_sash_ratio'] = ratio
                    self.tab_config['daily_usage_sash_pos'] = sash_pos
                    
                    self.save_tab_config()
                    print(f"Sash ratio saved: {ratio:.3f}")
        except Exception as e:
            print(f"Error saving sash position: {e}")

    def toggle_resolution_lock(self):
        """Toggle window resolution lock"""
        try:
            self.resolution_locked = not self.resolution_locked
            
            if self.resolution_locked:
                self.locked_width = self.root.winfo_width()
                self.locked_height = self.root.winfo_height()
                self.root.resizable(False, False)
                if hasattr(self, 'btn_resolution_lock'):
                    self.btn_resolution_lock.config(text="🔒 해상도 고정됨")
                
                if not hasattr(self, 'tab_config'):
                    self.tab_config = {}
                self.tab_config['resolution_locked'] = True
                self.tab_config['locked_width'] = self.locked_width
                self.tab_config['locked_height'] = self.locked_height
                print(f"Resolution locked at: {self.locked_width}x{self.locked_height}")
            else:
                self.root.resizable(True, True)
                if hasattr(self, 'btn_resolution_lock'):
                    self.btn_resolution_lock.config(text="🔓 해상도 고정")
                
                if not hasattr(self, 'tab_config'):
                    self.tab_config = {}
                self.tab_config['resolution_locked'] = False
                print("Resolution unlocked")
                
            self.save_tab_config(force=True)
        except Exception as e:
            print(f"Error toggling resolution lock: {e}")

    # Removed force_save_config - using save_tab_config(force=True) instead

    def extract_sn_from_model(self, model_name, current_sn):
        """Extract SN from model name if 'S/N.' is present and update current_sn if it's empty"""
        if pd.isna(model_name):
            return model_name, current_sn
        
        model_str = str(model_name)
        if 'S/N.' in model_str:
            parts = model_str.split('S/N.', 1)
            new_model = parts[0].strip()
            # Remove trailing underscore or dot if present before S/N.
            if new_model.endswith('_') or new_model.endswith('.'):
                new_model = new_model[:-1].strip()
            
            extracted_sn = parts[1].strip()
            
            # Use extracted SN if current SN is empty or matches part of it
            if not str(current_sn).strip() or str(current_sn) == 'nan':
                return new_model, extracted_sn
            else:
                # If current SN exists, we still clean up the model name
                return new_model, current_sn
        
        return model_name, current_sn

    def load_data(self):
        import re
        def normalize_cols(df):
            if df is not None and not df.empty:
                df.columns = [re.sub(r'\s+', '', str(c)) for c in df.columns]
            return df

        try:
            print(f"DEBUG: Loading data from {self.db_path}...")
            
            # Check if database exists in app_dir. If not, try to restore from bundle_dir
            if not os.path.exists(self.db_path):
                bundled_db = os.path.join(self.bundle_dir, 'Material_Inventory.xlsx')
                print(f"DEBUG: Main DB not found. Trying to restore from bundle: {bundled_db}")
                if os.path.exists(bundled_db):
                    import shutil
                    try:
                        shutil.copy2(bundled_db, self.db_path)
                        print("DEBUG: Restored DB from bundle.")
                        # Also try to copy config if it exists in bundle but not in app_dir
                        bundled_config = os.path.join(self.bundle_dir, 'Material_Manager_Config.json')
                        if os.path.exists(bundled_config) and not os.path.exists(self.config_path):
                            shutil.copy2(bundled_config, self.config_path)
                            print("DEBUG: Restored Config from bundle.")
                    except Exception as e:
                        print(f"Failed to restore data from bundle: {e}")

            if not os.path.exists(self.db_path):
                print("DEBUG: DB still not found. Initializing new DataFrames.")
                # Initialize with new schema if still not found
                self.materials_df = pd.DataFrame(columns=[
                    'MaterialID', '회사코드', '관리품번', '품목명', 'SN', '창고',
                    '모델명', '규격', '품목군코드', '공급업체', '제조사', '제조국', 
                    '가격', '원가', '관리단위', '수량', '재고하한', 'Active'
                ])
                self.transactions_df = pd.DataFrame(columns=['Date', 'MaterialID', 'Site', 'Type', 'Quantity', 'Note', 'User', '차량번호', '주행거리', '차량점검', '차량비고'])
                self.monthly_usage_df = pd.DataFrame(columns=['MaterialID', 'Year', 'Month', 'Site', 'Usage', 'Note', 'Entry Date'])
                self.daily_usage_df = pd.DataFrame(columns=['Date', 'Site', 'MaterialID', 'Usage', 'Note', 'EntryTime', 
                                                'RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크', 
                                                'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타', '장비명', '검사방법', '검사량',
                                                '단가', '출장비', '일식', '검사비', 'FilmCount',
                                                'User', 'WorkTime', 'OT',
                                                'User2', 'WorkTime2', 'OT2',
                                                'User3', 'WorkTime3', 'OT3',
                                                'User4', 'WorkTime4', 'OT4',
                                                'User5', 'WorkTime5', 'OT5',
                                                'User6', 'WorkTime6', 'OT6',
                                                'User7', 'WorkTime7', 'OT7',
                                                'User8', 'WorkTime8', 'OT8',
                                                'User9', 'WorkTime9', 'OT9',
                                                'User10', 'WorkTime10', 'OT10',
                                                '차량번호', '주행거리', '차량점검', '차량비고'])
                self.budget_df = pd.DataFrame(columns=['Site', 'Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit', 'Note', 'LaborDetail', 'MaterialDetail'])
            else:
                print("DEBUG: DB found. Reading Excel...")
                # 1. Materials
                self.materials_df = pd.read_excel(self.db_path, sheet_name='Materials')
                self.materials_df = normalize_cols(self.materials_df)
                
                # Handle column rename from '품명' to '품목명'
                if '품명' in self.materials_df.columns and '품목명' not in self.materials_df.columns:
                    self.materials_df.rename(columns={'품명': '품목명'}, inplace=True)
                
                # Add missing columns for backward compatibility
                missing_cols = {
                    'SN': '',
                    '공급업체': '',
                    '원가': 0,
                    '재고하한': 0,
                    'Active': True
                }
                for col, default in missing_cols.items():
                    if col not in self.materials_df.columns:
                        self.materials_df[col] = default
                
                # Force specific columns to string type and clean numeric artifacts (.0, -0.0)
                str_cols = ['회사코드', '관리품번', '품목명', 'SN', '창고', '모델명', '규격', 
                            '품목군코드', '공급업체', '제조사', '제조국', '관리단위']
                for col in str_cols:
                    if col in self.materials_df.columns:
                        # Ensure all strings are stripped and nan-free
                        self.materials_df[col] = self.materials_df[col].astype(str).str.strip().replace(['nan', 'None', 'NULL', '-0.0', '0.0', 'NaN', 'NaN.0'], '')
                        self.materials_df[col] = self.materials_df[col].str.replace(r'\.0$', '', regex=True)
                
                # 2. Transactions
                self.transactions_df = pd.read_excel(self.db_path, sheet_name='Transactions')
                self.transactions_df = normalize_cols(self.transactions_df)
                
                # Ensure it has all required columns
                for col in ['Site', 'User', 'Note', 'MaterialID', 'Type', 'Quantity', 'Date', '차량번호', '주행거리', '차량점검', '차량비고']:
                    if col not in self.transactions_df.columns:
                        self.transactions_df[col] = '' if col != 'Quantity' else 0.0
                
                # Ensure Date column is datetime and MaterialID is numeric
                if not self.transactions_df.empty:
                    self.transactions_df['Date'] = pd.to_datetime(self.transactions_df['Date'], errors='coerce')
                    self.transactions_df['MaterialID'] = pd.to_numeric(self.transactions_df['MaterialID'], errors='coerce')
                    
                    # Force string columns and clean numeric artifacts
                    for col in ['Type', 'Note', 'User', 'Site', '차량번호', '주행거리', '차량점검', '차량비고']:
                        self.transactions_df[col] = self.transactions_df[col].astype(str).replace(['nan', 'None', 'NULL', '-0.0', '0.0', 'NaN'], '')
                        self.transactions_df[col] = self.transactions_df[col].str.replace(r'\.0$', '', regex=True)
                    
                    # One-time cleanup: Remove '현장사용' and redundant model names from historical notes
                    self.transactions_df['Note'] = self.transactions_df['Note'].astype(str).str.replace('현장사용', '', regex=False).str.strip()
                    
                    # Clean up notes that are identical to model names
                    if not self.transactions_df.empty and not self.materials_df.empty:
                        # Create a map for MaterialID -> Model Name
                        id_to_model = self.materials_df.set_index('MaterialID')['모델명'].astype(str).to_dict()
                        
                        def clean_redundant_note(row):
                            note = str(row['Note']).strip()
                            mat_id = row['MaterialID']
                            model = str(id_to_model.get(mat_id, '')).strip()
                            if note and model and note == model:
                                return ''
                            return note
                            
                        self.transactions_df['Note'] = self.transactions_df.apply(clean_redundant_note, axis=1)
                    
                    # Normalize OUT transaction quantities to negative and strip Site names
                    if not self.transactions_df.empty:
                        if 'Site' in self.transactions_df.columns:
                            self.transactions_df['Site'] = self.transactions_df['Site'].astype(str).str.strip().replace(['nan', 'None'], '')
                        
                        out_mask = (self.transactions_df['Type'] == 'OUT') & (self.transactions_df['Quantity'] > 0)
                        if out_mask.any():
                            self.transactions_df.loc[out_mask, 'Quantity'] = -self.transactions_df.loc[out_mask, 'Quantity']
                            print(f"DEBUG: Normalized {out_mask.sum()} OUT transactions to negative.")
                
                # 3. Monthly Usage
                try:
                    self.monthly_usage_df = pd.read_excel(self.db_path, sheet_name='MonthlyUsage', dtype={'Site': str, 'Note': str})
                    self.monthly_usage_df = normalize_cols(self.monthly_usage_df)
                    
                    if not self.monthly_usage_df.empty:
                        self.monthly_usage_df['MaterialID'] = pd.to_numeric(self.monthly_usage_df['MaterialID'], errors='coerce')
                        self.monthly_usage_df['EntryDate'] = pd.to_datetime(self.monthly_usage_df['EntryDate'])
                        # Add Site column if it doesn't exist (for backward compatibility) and strip
                        if 'Site' not in self.monthly_usage_df.columns:
                            self.monthly_usage_df['Site'] = ''
                        else:
                            self.monthly_usage_df['Site'] = self.monthly_usage_df['Site'].astype(str).str.strip().replace(['nan', 'None'], '')
                except Exception as e:
                    print(f"DEBUG: Failed to load MonthlyUsage: {e}")
                    self.monthly_usage_df = pd.DataFrame(columns=['MaterialID', 'Year', 'Month', 'Site', 'Usage', 'Note', 'EntryDate'])
                
                # 4. Daily Usage
                try:
                    # Explicitly set dtypes to avoid float inference for empty columns
                    self.daily_usage_df = pd.read_excel(self.db_path, sheet_name='DailyUsage', 
                                                        dtype={'Site': str, 'Note': str, 'User': str})
                    self.daily_usage_df = normalize_cols(self.daily_usage_df)
                                                        
                    if not self.daily_usage_df.empty:
                        self.daily_usage_df['Date'] = pd.to_datetime(self.daily_usage_df['Date'])
                        self.daily_usage_df['EntryTime'] = pd.to_datetime(self.daily_usage_df['EntryTime'])
                        
                        # Fill NaNs and clean numeric artifacts in string columns
                        string_columns = ['Site', 'Note', '장비명', '검사방법']
                        # Add all users, worktimes and OT columns
                        for i in range(1, 11):
                            u_col = 'User' if i == 1 else f'User{i}'
                            w_col = 'WorkTime' if i == 1 else f'WorkTime{i}'
                            o_col = 'OT' if i == 1 else f'OT{i}'
                            string_columns.extend([u_col, w_col, o_col])
                        
                        # Add vehicle columns
                        string_columns.extend(['차량번호', '주행거리', '차량점검', '차량비고'])
                            
                        for col in string_columns:
                            if col not in self.daily_usage_df.columns:
                                self.daily_usage_df[col] = ''
                            # Ensure all strings are stripped and nan-free
                            self.daily_usage_df[col] = self.daily_usage_df[col].astype(str).str.strip().replace(['nan', 'None', 'NULL', '-0.0', '0.0', 'NaN', 'NAN', 'nan.0'], '')
                        
                        # Add/Fix columns for auto-calculation
                        if '검사방법' not in self.daily_usage_df.columns and '검사량' in self.daily_usage_df.columns:
                            # Migrate old string '검사량' (PAUT, UT...) to '검사방법'
                            self.daily_usage_df['검사방법'] = self.daily_usage_df['검사량'].astype(str)
                            self.daily_usage_df['검사량'] = 0.0
                        
                        for col in ['검사량', '단가', '출장비', '일식', '검사비', '수량']:
                            if col not in self.daily_usage_df.columns:
                                self.daily_usage_df[col] = 0.0
                            else:
                                self.daily_usage_df[col] = pd.to_numeric(self.daily_usage_df[col], errors='coerce').fillna(0.0)
                        
                        # [NEW] Compatibility: If '수량' exists but '검사량' is empty/missing, map '수량' to '검사량'
                        if '수량' in self.daily_usage_df.columns:
                             # If both exist, we prefer '수량' if '검사량' is all zeros or NaNs
                             if '검사량' in self.daily_usage_df.columns:
                                 mask = (self.daily_usage_df['검사량'] == 0) | (self.daily_usage_df['검사량'].isna())
                                 self.daily_usage_df.loc[mask, '검사량'] = self.daily_usage_df.loc[mask, '수량']
                             else:
                                 self.daily_usage_df['검사량'] = self.daily_usage_df['수량']
                        
                        # Ensure MaterialID is numeric
                        if 'MaterialID' in self.daily_usage_df.columns:
                            self.daily_usage_df['MaterialID'] = pd.to_numeric(self.daily_usage_df['MaterialID'], errors='coerce')
                          # Add RTK columns if they don't exist (for backward compatibility)
                        rtk_columns = ['RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크', 
                                      'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타']
                        for col in rtk_columns:
                            if col not in self.daily_usage_df.columns:
                                self.daily_usage_df[col] = 0.0
                        # Remove old RTK Category column if exists
                        if 'RTK Category' in self.daily_usage_df.columns:
                            self.daily_usage_df = self.daily_usage_df.drop('RTK Category', axis=1)
                except Exception as e:
                    print(f"DEBUG: Failed to load DailyUsage: {e}")
                    self.daily_usage_df = pd.DataFrame(columns=['Date', 'Site', 'MaterialID', 'Usage', 'Note', 'EntryTime', 
                                                        'RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크', 
                                                        'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타', 'User', '장비명', '검사량',
                                                        '단가', '출장비', '일식', '검사비',
                                                        '차량번호', '주행거리', '차량점검', '차량비고'])
                
                # 5. Migrate vehicle data from daily_usage_df to transactions_df
                if hasattr(self, 'daily_usage_df') and hasattr(self, 'transactions_df'):
                    if not self.daily_usage_df.empty and not self.transactions_df.empty:
                        # Group daily usage by date, site, material to match transactions
                        for _, daily_row in self.daily_usage_df.iterrows():
                            # Handle NaT values safely
                            daily_date = pd.to_datetime(daily_row['Date'], errors='coerce')
                            if pd.isna(daily_date):
                                continue  # Skip rows with invalid dates
                            daily_date = daily_date.normalize()
                            daily_site = str(daily_row['Site']).strip()
                            daily_mat = daily_row['MaterialID']
                            
                            # Find matching transactions
                            mask = (
                                (pd.to_datetime(self.transactions_df['Date'], errors='coerce').dt.normalize() == daily_date) &
                                (self.transactions_df['Site'].str.strip() == daily_site) &
                                (self.transactions_df['MaterialID'] == daily_mat)
                            )
                            
                            # Update vehicle info for matching transactions
                            if mask.any():
                                self.transactions_df.loc[mask, '차량번호'] = str(daily_row.get('차량번호', ''))
                                self.transactions_df.loc[mask, '주행거리'] = str(daily_row.get('주행거리', ''))
                                self.transactions_df.loc[mask, '차량점검'] = str(daily_row.get('차량점검', ''))
                                self.transactions_df.loc[mask, '차량비고'] = str(daily_row.get('차량비고', ''))


                
                # Add SN column if it doesn't exist (for backward compatibility)
                if 'SN' not in self.materials_df.columns:
                    self.materials_df['SN'] = ''
                
                # Migrate old schema if needed
                if 'Equipment Code' in self.materials_df.columns and '회사코드' not in self.materials_df.columns:
                    self.migrate_old_schema()
                
                # Apply SN extraction from Model Name to existing data
                if not self.materials_df.empty:
                    updated = False
                    for idx, row in self.materials_df.iterrows():
                        model = row.get('모델명', '')
                        sn = row.get('SN', '')
                        new_model, new_sn = self.extract_sn_from_model(model, sn)
                        
                        if str(model) != str(new_model) or str(sn) != str(new_sn):
                            self.materials_df.at[idx, '모델명'] = new_model
                            self.materials_df.at[idx, 'SN'] = new_sn
                            updated = True
                    
                    if updated:
                        self.save_data()

                # 6. Budget
                try:
                    self.budget_df = pd.read_excel(self.db_path, sheet_name='Budget')
                    self.budget_df = normalize_cols(self.budget_df)
                    # Add missing detail columns for backward compatibility
                    if 'UnitPrice' not in self.budget_df.columns:
                        self.budget_df['UnitPrice'] = ''
                    if 'LaborDetail' not in self.budget_df.columns:
                        self.budget_df['LaborDetail'] = ''
                    if 'MaterialDetail' not in self.budget_df.columns:
                        self.budget_df['MaterialDetail'] = ''
                    if 'ExpenseDetail' not in self.budget_df.columns:
                        self.budget_df['ExpenseDetail'] = ''
                except Exception as e:
                    print(f"DEBUG: Budget sheet not found or load failed: {e}")
                    self.budget_df = pd.DataFrame(columns=['Site', 'Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit', 'Note', 'LaborDetail', 'MaterialDetail', 'ExpenseDetail'])
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.show_error_dialog("Error", f"데이터를 불러오는데 실패했습니다: {e}")
            self.materials_df = pd.DataFrame(columns=[
                'MaterialID', '회사코드', '관리품번', '품목명', 'SN', '창고',
                '모델명', '규격', '품목군코드', '공급업체', '제조사', '제조국', 
                '가격', '관리단위', '수량', '재고하한'
            ])
            self.transactions_df = pd.DataFrame(columns=['Date', 'MaterialID', 'Site', 'Type', 'Quantity', 'Note', 'User', '차량번호', '주행거리', '차량점검', '차량비고'])
            self.monthly_usage_df = pd.DataFrame(columns=['MaterialID', 'Year', 'Month', 'Site', 'Usage', 'Note', 'Entry Date'])
            self.daily_usage_df = pd.DataFrame(columns=['Date', 'Site', 'MaterialID', 'Usage', 'Note', 'EntryTime', 
                                                'RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크', 
                                                'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타', '장비명', '검사량',
                                            '차량번호', '주행거리', '차량점검', '차량비고'])
            self.budget_df = pd.DataFrame(columns=['Site', 'Revenue', 'UnitPrice', 'LaborCost', 'MaterialCost', 'Expense', 'OutsourceCost', 'Profit', 'Note'])
        
        # Global header cleanup: remove ALL internal and edge whitespace for permanent stability
        for df_attr in ['materials_df', 'transactions_df', 'daily_usage_df', 'budget_df']:
            if hasattr(self, df_attr):
                df = getattr(self, df_attr)
                if df is not None and not df.empty:
                    df.columns = [re.sub(r'\s+', '', str(c)) for c in df.columns]
                    setattr(self, df_attr, df)
            
            # Refresh all inquiry filters once data is loaded
            self.refresh_inquiry_filters()

    
    def migrate_old_schema(self):
        """Migrate data from old schema to new schema"""
        old_df = self.materials_df.copy()
        self.materials_df = pd.DataFrame(columns=[
            'MaterialID', '회사코드', '관리품번', '품목명', 'SN', '창고',
            '모델명', '규격', '품목군코드', '제조사', '제조국', 
            '가격', '원가', '관리단위', '수량'
        ])
        
        # Headers are already cleaned by the global cleanup in load_data/init

        for _, row in old_df.iterrows():
            new_row = {
                'MaterialID': row.get('MaterialID', ''),
                '회사코드': '',
                '관리품번': row.get('Equipment Code', ''),
                '품목명': row.get('Item Name', row.get('Name', '')),
                'SN': row.get('SN', ''),
                '창고': '',
                '모델명': '',
                '규격': row.get('Specification', ''),
                '품목군코드': '',
                '제조사': row.get('Manufacturer', ''),
                '제조국': '',
                '가격': 0,
                '관리단위': row.get('Unit', 'EA'),
                '수량': row.get('Current Stock', row.get('Initial Stock', 0))
            }
            self.materials_df = pd.concat([self.materials_df, pd.DataFrame([new_row])], ignore_index=True)
        
        self.save_data()
        messagebox.showinfo("마이그레이션 완료", "기존 데이터가 새로운 형식으로 변환되었습니다.")
    
    def enable_autocomplete(self, combobox, values_list_attr=None, values_list=None, prefix_list=None):
        """
        Enable autocomplete/autosuggestion on a Combobox widget with a custom Toplevel dropdown.
        This keeps the cursor (caret) 100% visible and stable.
        """
        if not hasattr(self, '_autocomplete_timers'):
            self._autocomplete_timers = {}
        if not hasattr(self, '_suggestion_windows'):
            self._suggestion_windows = {}

        # Create/Get custom suggestion window for this combobox
        if combobox not in self._suggestion_windows:
            self._suggestion_windows[combobox] = SuggestionWindow(combobox)
        
        suggest_win = self._suggestion_windows[combobox]

        def perform_filter(event_widget, force_all=False):
            typed = str(event_widget.get()).strip() # [NEW] Strip input for robustness
            
            if values_list_attr:
                base_values = getattr(self, values_list_attr, [])
            elif values_list is not None:
                base_values = values_list
            else:
                base_values = []
            
            # Combine with prefix list (e.g. ['전체']) if provided
            all_values = (prefix_list + base_values) if prefix_list else base_values
            
            # [FIX] If forced (click/focus), show all values even if something is typed/selected
            if force_all or not typed:
                event_widget['values'] = all_values
                suggest_win.show(all_values)
                return
            
            # [NEW] If matches prefix_list (like "전체"), show all options
            is_prefix = (prefix_list and typed in prefix_list)
            if is_prefix:
                event_widget['values'] = all_values
                suggest_win.show(all_values)
                return
            
            filtered = [v for v in all_values if str(v).lower().startswith(typed.lower())]
            event_widget['values'] = filtered
            
            # Show custom window
            suggest_win.show(filtered)

        def on_keyrelease(event):
            # Handle navigation keys for custom window
            if event.keysym == "Down" and suggest_win.active:
                suggest_win.move_selection(1)
                return "break"
            if event.keysym == "Up" and suggest_win.active:
                suggest_win.move_selection(-1)
                return "break"
            if event.keysym == "Escape":
                suggest_win.hide()
                return "break"

            if event.keysym in ("Left", "Right", "Up", "Down", "Return", "Escape", "Tab", "Shift_L", "Shift_R", "Control_L", "Control_R"):
                if event.keysym in ("Escape", "Tab"):
                    suggest_win.hide()
                return
            
            widget = event.widget
            if widget in self._autocomplete_timers:
                self.root.after_cancel(self._autocomplete_timers[widget])
            
            timer_id = self.root.after(50, lambda: perform_filter(widget, force_all=False))
            self._autocomplete_timers[widget] = timer_id
        
        combobox.bind('<KeyRelease>', on_keyrelease, add=True)
        
        def handle_extra_nav(event):
            if not suggest_win.active: return
            if event.keysym == "Down": suggest_win.move_selection(1); return "break"
            if event.keysym == "Up": suggest_win.move_selection(-1); return "break"
            if event.keysym == "Return": 
                if suggest_win.confirm_selection(): return "break"
                
        combobox.bind('<KeyPress-Down>', handle_extra_nav, add=True)
        combobox.bind('<KeyPress-Up>', handle_extra_nav, add=True)
        combobox.bind('<KeyPress-Return>', handle_extra_nav, add=True)
        combobox.bind('<KeyPress-Escape>', lambda e: suggest_win.hide(), add=True)
        combobox.bind('<KeyPress-Tab>', lambda e: suggest_win.hide(), add=True)

        def on_focus_out(event):
            self.root.after(200, suggest_win.hide)
        # Use add=True to preserve external save bindings
        combobox.bind('<FocusOut>', on_focus_out, add=True)
        combobox.bind('<<ComboboxSelected>>', lambda e: suggest_win.hide(), add=True)

        def on_focus_in(event):
            # [NEW] Select all text for easy overwriting
            try:
                event.widget.selection_range(0, tk.END)
            except: pass
            # [FIX] Trigger filter with force_all=True on focus
            perform_filter(event.widget, force_all=True)
        combobox.bind('<FocusIn>', on_focus_in, add=True)
        
        # [FIX] Also show list on click to handle cases where it's already focused
        def on_click(event):
            perform_filter(event.widget, force_all=True)
        combobox.bind('<Button-1>', on_click, add=True)


    def apply_autocomplete_to_all_comboboxes(self):
        """Map specific comboboxes to their data lists for autocomplete"""
        # Mapping: {widget_attr_name: values_list_attr_name}
        mappings = {
            'cb_daily_site': 'sites',
            'cb_daily_equip': 'equipment_suggestions',
            'cb_material': 'materials_display_list',
            'cb_daily_material': 'materials_display_list',
            'cb_trans_filter_mat': 'materials_display_list',
            'cb_trans_filter_site': 'sites',
            'cb_daily_filter_site': 'sites',
            'cb_budget_site': 'budget_sites',
            'cb_daily_filter_material': 'materials_display_list',
            'cb_daily_filter_worker': 'users',
            'cb_daily_filter_vehicle': 'vehicles',
            'cb_trans_filter_vehicle': 'vehicles',
            'cb_sales_filter_site': 'sites',
            'cb_filter_co': 'co_code_list',
            'cb_filter_class': 'class_list',
            'cb_filter_mfr': 'mfr_list',
            'cb_filter_name': 'materials_display_list',
            'cb_filter_sn': 'sn_list',
            'cb_filter_model': 'model_list',
            'cb_filter_eq': 'eq_code_list',
            'cb_co_code': 'co_code_list',
            'cb_eq_code': 'eq_code_list',
            'cb_item_name': 'item_name_list',
            'cb_class': 'class_list',
            'cb_spec': 'spec_list',
            'cb_unit': 'unit_list',
            'cb_mfr': 'mfr_list',
            'cb_origin': 'origin_list'
        }
        
        for widget_attr, list_attr in mappings.items():
            if hasattr(self, widget_attr):
                widget = getattr(self, widget_attr)
                if isinstance(widget, ttk.Combobox):
                    # For filtering, we might want to include '전체' dynamically
                    current_values = list(widget['values'])
                    if '전체' in current_values:
                        self.enable_autocomplete(widget, values_list_attr=list_attr, prefix_list=['전체'])
                    else:
                        self.enable_autocomplete(widget, values_list_attr=list_attr)

        # Handle cb_daily_test_method with static list
        if hasattr(self, 'cb_daily_test_method'):
            self.enable_autocomplete(self.cb_daily_test_method, values_list=["RT", "PAUT", "UT", "MT", "PT", "PMI"])

    def _safe_format_datetime(self, val, format_str='%Y-%m-%d %H:%M'):
        if pd.isna(val) or val is None or str(val).strip().lower() in ['nan', 'none', '']:
            return ""
        try:
            return pd.to_datetime(val).strftime(format_str)
        except:
            return str(val)

    def clean_nan(self, val):
        """Robustly clean NaN, None, and other empty markers from any value"""
        if pd.isna(val) or val is None: return ""
        s = str(val).strip()
        if not s or NAN_PATTERN.match(s):
            return ""
        # Handle trailing .0 from numeric inputs converted to string
        s = DOT_ZERO_PATTERN.sub('', s)
        if NAN_PATTERN.match(s): return ""
        return s

    def save_data(self):
        try:
            
            # [STABILITY] Ensure MaterialID is consistently numeric in all sheets
            # Only cast if types are mixed or object to save time
            for df_name, df in [('Materials', self.materials_df), ('Transactions', self.transactions_df), 
                                ('Monthly', self.monthly_usage_df), ('Daily', self.daily_usage_df)]:
                if df is not None and 'MaterialID' in df.columns:
                    # Optimized check: only cast if needed
                    if df['MaterialID'].dtype == object:
                        df['MaterialID'] = pd.to_numeric(df['MaterialID'], errors='coerce')
            
            # Explicitly check for write permission/locks
            if os.path.exists(self.db_path):
                try:
                    with open(self.db_path, 'a'): pass
                except PermissionError:
                    raise Exception(f"File '{os.path.basename(self.db_path)}' is currently open in another program (like Excel). Please close it and try again.")

            with pd.ExcelWriter(self.db_path, engine='openpyxl') as writer:
                self.materials_df.to_excel(writer, sheet_name='Materials', index=False)
                self.transactions_df.to_excel(writer, sheet_name='Transactions', index=False)
                self.monthly_usage_df.to_excel(writer, sheet_name='MonthlyUsage', index=False)
                self.daily_usage_df.to_excel(writer, sheet_name='DailyUsage', index=False)
                self.budget_df.to_excel(writer, sheet_name='Budget', index=False)
            
            return True
        except Exception as e:
            import traceback
            err_detail = traceback.format_exc()
            self.show_error_dialog("데이터 저장 실패", f"데이터를 저장하는데 실패했습니다:\n{e}\n\n파일이 열려있다면 닫고 다시 시도해주세요.")
            return False

    def create_widgets(self):
        # Notebook for Tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill='both', padx=10, pady=10)
        
        # Tab 1: Current Stock
        self.tab_stock = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_stock, text='현재 재고 현황')
        self.setup_stock_tab()
        
        # Tab 2: Register/Transaction
        self.tab_inout = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_inout, text='입출고 관리')
        self.setup_inout_tab()
        

        # Tab 4: Import/Export
        self.tab_import = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_import, text='데이터 가져오기/내보내기')
        self.setup_import_tab()
        
        # Tab 5: Monthly Usage Entry
        self.tab_monthly_usage = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_monthly_usage, text='월별 집계')
        self.setup_monthly_usage_tab()
        
        # Tab 6: Daily Usage Entry by Site
        self.tab_daily_usage = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_daily_usage, text='현장별 일일 사용량 기입')
        self.setup_daily_usage_tab()
        

        # Tab 8: Project Execution Budget (New)
        self.tab_budget = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_budget, text='공사실행예산서')
        self.setup_budget_tab()
        
        # Initial view update (Move here to ensure all tabs are ready before refresh)
        self.refresh_inquiry_filters()
        self.update_daily_usage_view()
        
        # Bind tab events (Drag and drop reordering)
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        self.notebook.bind("<Button-1>", self.on_tab_drag_start)
        self.notebook.bind("<B1-Motion>", self.on_tab_drag)
        self.notebook.bind("<ButtonRelease-1>", self.on_tab_drag_end)
        self.notebook.bind("<Button-3>", self.show_tab_context_menu)
        
        # Bind main window resize to maintain sash ratios
        self.root.bind("<Configure>", self._on_main_window_resize)
        
        # Load and restore all configuration (geometry, locks, sashes, draggable items)
        self.root.after(100, self.load_tab_config)
        
        # Apply autocomplete to all comboboxes
        self.root.after(200, self.apply_autocomplete_to_all_comboboxes)
        
        # Save tab config on window close
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    def setup_stock_tab(self):
        # Control Frame (Vertical container for both rows)
        control_frame = ttk.Frame(self.tab_stock)
        control_frame.pack(fill='x', padx=5, pady=5)
        
        # Row 1: Action Buttons
        action_row = ttk.Frame(control_frame)
        action_row.pack(fill='x', side='top', pady=(0, 5))
        
        btn_refresh = ttk.Button(action_row, text="재고 새로고침", command=self.update_stock_view)
        btn_refresh.pack(side='left', padx=5)
        
        btn_alert = ttk.Button(action_row, text="재주문 필요 항목 보기", command=self.show_low_stock)
        btn_alert.pack(side='left', padx=5)
        
        btn_delete = ttk.Button(action_row, text="품목 삭제", command=self.delete_selected_material)
        btn_delete.pack(side='left', padx=5)
        
        btn_edit = ttk.Button(action_row, text="품목 수정", command=self.open_edit_material_dialog)
        btn_edit.pack(side='left', padx=5)
        
        btn_export = ttk.Button(action_row, text="엑셀 내보내기", command=self.export_stock_to_excel)
        btn_export.pack(side='left', padx=5)
        
        btn_select_all = ttk.Button(action_row, text="전체 선택", command=self.select_all_stock)
        btn_select_all.pack(side='left', padx=5)
        
        # Row 2: Search and Filter Frame
        filter_row = ttk.Frame(control_frame)
        filter_row.pack(fill='x', side='top')
        
        filter_frame = ttk.LabelFrame(filter_row, text="검색 필터")
        filter_frame.pack(fill='x', padx=5, pady=2)
        
        # Row 0 of Filter Frame (Grid)
        ttk.Label(filter_frame, text="회사:").grid(row=0, column=0, padx=2, pady=2, sticky='e')
        self.cb_filter_co = ttk.Combobox(filter_frame, width=15)
        self.cb_filter_co.grid(row=0, column=1, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="분류:").grid(row=0, column=2, padx=2, pady=2, sticky='e')
        self.cb_filter_class = ttk.Combobox(filter_frame, width=15)
        self.cb_filter_class.grid(row=0, column=3, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="제조사:").grid(row=0, column=4, padx=2, pady=2, sticky='e')
        self.cb_filter_mfr = ttk.Combobox(filter_frame, width=15)
        self.cb_filter_mfr.grid(row=0, column=5, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="품목명:").grid(row=0, column=6, padx=2, pady=2, sticky='e')
        self.cb_filter_name = ttk.Combobox(filter_frame, width=25)
        self.cb_filter_name.grid(row=0, column=7, padx=2, pady=2)
        
        # Row 1 of Filter Frame
        ttk.Label(filter_frame, text="S/N:").grid(row=1, column=0, padx=2, pady=2, sticky='e')
        self.cb_filter_sn = ttk.Combobox(filter_frame, width=20)
        self.cb_filter_sn.grid(row=1, column=1, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="모델명:").grid(row=1, column=2, padx=2, pady=2, sticky='e')
        self.cb_filter_model = ttk.Combobox(filter_frame, width=20)
        self.cb_filter_model.grid(row=1, column=3, padx=2, pady=2)
        
        ttk.Label(filter_frame, text="관리품번:").grid(row=1, column=4, padx=2, pady=2, sticky='e')
        self.cb_filter_eq = ttk.Combobox(filter_frame, width=20)
        self.cb_filter_eq.grid(row=1, column=5, padx=2, pady=2)

        # [NEW] Bind actions to all stock filter comboboxes for consistency and focus handling
        stock_filters = [
            self.cb_filter_co, self.cb_filter_class, self.cb_filter_mfr,
            self.cb_filter_name, self.cb_filter_sn, self.cb_filter_model,
            self.cb_filter_eq
        ]

        def on_stock_filter_action(e):
            self.update_stock_view()
            # [NEW] Move focus to the next filter in sequence for "sideways" navigation
            try:
                current_idx = stock_filters.index(e.widget)
                if current_idx + 1 < len(stock_filters):
                    stock_filters[current_idx + 1].focus_set()
                else:
                    # Move to the General Search entry after all dropdowns
                    self.search_entry.focus_set()
            except:
                # Fallback to result list if sequence fails
                self.stock_tree.focus_set()

        for combo in stock_filters:
            combo.bind("<Return>", on_stock_filter_action)
            combo.bind('<<ComboboxSelected>>', lambda e: self.update_stock_view())
        
        ttk.Label(filter_frame, text="검색어:").grid(row=1, column=6, padx=2, pady=2, sticky='e')
        self.search_var = tk.StringVar()
        self.search_var.trace_add('write', lambda *args: self.update_stock_view())
        self.search_entry = ttk.Entry(filter_frame, textvariable=self.search_var, width=20)
        self.search_entry.grid(row=1, column=7, padx=2, pady=2)
        
        # [NEW] Special handler for Search Entry to move to Results List
        def on_search_enter(e):
            self.update_stock_view()
            self.stock_tree.focus_set()
            if self.stock_tree.get_children():
                # Highlight first item for immediate keyboard control
                self.stock_tree.selection_set(self.stock_tree.get_children()[0])
        self.search_entry.bind("<Return>", on_search_enter)
        
        # Reset Filters Button
        btn_reset = ttk.Button(filter_frame, text="필터 초기화", command=self.reset_stock_filters)
        btn_reset.grid(row=1, column=8, padx=10, pady=2)
        
        # Treeview for Stock with Scrollbars
        tree_frame = ttk.Frame(self.tab_stock)
        tree_frame.pack(expand=True, fill='both', padx=5, pady=5)
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        columns = ('ID', '회사코드', '관리품번', '품목명', 'SN', '창고', '모델명', '규격', '품목군코드', '공급업체', '제조사', '제조국', '가격', '원가', '관리단위', '수량', '재고하한')
        self.stock_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', 
                                      yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        vsb.config(command=self.stock_tree.yview)
        hsb.config(command=self.stock_tree.xview)
        
        # Column configuration
        col_widths = [60, 80, 100, 180, 90, 90, 120, 120, 90, 120, 120, 80, 80, 80, 80, 80, 80]
        for col, width in zip(columns, col_widths):
            self.stock_tree.heading(col, text=col)
            # Change stretch=True to stretch=False to allow fixed user-defined widths
            self.stock_tree.column(col, width=width, minwidth=50, stretch=False, anchor='center')
        
        # Bind double-click
        self.stock_tree.bind('<Double-1>', lambda e: self.open_edit_material_dialog())
        
        # Auto-save column widths when user resizes columns
        self.stock_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())
        
        # Grid layout
        self.stock_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Initial view update will be triggered after update_registration_combos 
        # to ensure filters are properly initialized to "전체" first.
    
    def show_low_stock(self):
        """Show items with low stock (less than their specific reorder point)"""
        low_stock_items = []
        for _, mat in self.materials_df.iterrows():
            if mat.get('Active', True) == False:
                continue
                
            current = self.calculate_current_stock(mat['MaterialID'])
            reorder_point = mat.get('재고하한', 10)
            if pd.isna(reorder_point) or reorder_point <= 0:
                reorder_point = 10 # Default fallback
                
            if pd.notna(current) and current < reorder_point:
                low_stock_items.append((mat.get('품목명', ''), current, reorder_point))
        
        if not low_stock_items:
            messagebox.showinfo("재고 알림", "수량이 10개 미만인 항목이 없습니다.")
        else:
            msg = "다음 항목들의 재고가 부족합니다:\n\n"
            for item, current, reorder in low_stock_items:
                msg += f"• {item}: 현재 {current:g} (필요 수준: {reorder:g})\n"
            messagebox.showwarning("재고 부족", msg)
    
    def select_all_stock(self):
        """Select all items in the stock treeview"""
        all_items = self.stock_tree.get_children()
        self.stock_tree.selection_set(all_items)

    def delete_selected_material(self):
        """선택된 자재 항목들을 재고에서 삭제"""
        selected_items = self.stock_tree.selection()
        
        if not selected_items:
            messagebox.showwarning("선택 오류", "삭제할 품목을 선택해주세요.")
            return
        
        # Confirm deletion
        confirm = messagebox.askyesno("삭제 확인", f"선택한 {len(selected_items)}개의 품목을 재고에서 영구히 삭제하시겠습니까?\n이 작업은 되돌릴 수 없습니다.")
        
        if not confirm:
            return
            
        # Get MaterialIDs to delete
        mat_ids_to_remove = []
        for item in selected_items:
            values = self.stock_tree.item(item, 'values')
            if values:
                # Ensure we match the type of MaterialID in the dataframe
                mat_ids_to_remove.append(type(self.materials_df['MaterialID'].iloc[0])(values[0]))
        
        # Soft delete: Set Active=False instead of removing from materials_df
        initial_count = len(self.materials_df[self.materials_df['Active'] != False])
        
        # We need to handle potential 'Active' column not existing gracefully just in case
        if 'Active' not in self.materials_df.columns:
            self.materials_df['Active'] = True
            
        self.materials_df.loc[self.materials_df['MaterialID'].isin(mat_ids_to_remove), 'Active'] = False
        
        removed_count = initial_count - len(self.materials_df[self.materials_df['Active'] != False])
        
        if removed_count > 0:
            # Save data and update views
            self.save_data()
            self.update_stock_view()
            self.update_material_combo()
            
            # Optional: Clear transactions related to these materials?
            # For now, let's keep them for history unless explicitly asked.
            
            messagebox.showinfo("완료", f"{removed_count}개의 품목이 삭제되었습니다.")
        else:
            messagebox.showwarning("실패", "데이터프레임에서 항목을 삭제하지 못했습니다.")

    def reset_stock_filters(self):
        """Reset all stock filters to default"""
        self.cb_filter_co.set("전체")
        self.cb_filter_class.set("전체")
        self.cb_filter_mfr.set("전체")
        self.cb_filter_name.set("전체")
        self.cb_filter_sn.set("전체")
        self.cb_filter_model.set("전체")
        self.cb_filter_eq.set("전체")
        self.search_var.set("")
        self.update_stock_view()

    def open_edit_material_dialog(self):
        """Open a dialog to edit the selected material"""
        selection = self.stock_tree.selection()
        if not selection:
            messagebox.showwarning("선택 오류", "수정할 자재를 선택해주세요.")
            return
            
        item = self.stock_tree.item(selection[0])
        mat_values = item['values']
        mat_id = mat_values[0]
        
        # Get full material data from DF
        mat_data = self.materials_df[self.materials_df['MaterialID'] == mat_id].iloc[0]
        
        # Create Edit Dialog
        edit_win = tk.Toplevel(self.root)
        edit_win.title("자재 정보 수정")
        edit_win.geometry("500x600")
        edit_win.transient(self.root)
        edit_win.grab_set()
        
        main_frame = ttk.Frame(edit_win, padding=20)
        main_frame.pack(expand=True, fill='both')
        
        fields = [
            ('회사코드', '회사코드'),
            ('관리품번', '관리품번'),
            ('품목명', '품목명'),
            ('SN', 'SN'),
            ('창고', '창고'),
            ('모델명', '모델명'),
            ('규격', '규격'),
            ('품목군코드', '품목군코드'),
            ('공급업체', '공급업체'),
            ('제조사', '제조사'),
            ('제조국', '제조국'),
            ('가격', '가격'),
            ('원가', '원가'),
            ('관리단위', '관리단위'),
            ('수량', '수량'),
            ('재고하한', '재고하한')
        ]
        
        entries = {}
        for i, (label_text, col_name) in enumerate(fields):
            ttk.Label(main_frame, text=f"{label_text}:").grid(row=i, column=0, padx=5, pady=5, sticky='w')
            
            # Using Combobox for some fields to maintain consistency
            if col_name in ['회사코드', '품목군코드', '제조사', '제조국', '관리단위']:
                ent = ttk.Combobox(main_frame, width=35)
                # Populate values from existing data
                if col_name in self.materials_df.columns:
                    unique_vals = sorted(self.materials_df[col_name].dropna().unique().tolist())
                    ent['values'] = unique_vals
            else:
                ent = ttk.Entry(main_frame, width=38)
                
            ent.grid(row=i, column=1, padx=5, pady=5)
            
            # Pre-fill value
            val = mat_data.get(col_name, '')
            if pd.isna(val): val = ''
            ent.insert(0, str(val))
            entries[col_name] = ent
            
        def on_save():
            new_data = {col: ent.get() for col, ent in entries.items()}
            self.save_material_edits(mat_id, new_data)
            edit_win.destroy()
            
        btn_save = ttk.Button(main_frame, text="변경사항 저장", command=on_save)
        btn_save.grid(row=len(fields), column=0, columnspan=2, pady=20)

    def save_material_edits(self, mat_id, new_data):
        """Save edited material data back to the database"""
        # Update DataFrame
        idx = self.materials_df.index[self.materials_df['MaterialID'] == mat_id].tolist()
        if not idx:
            messagebox.showerror("오류", "자재를 찾을 수 없습니다.")
            return
            
        for col, val in new_data.items():
            if col in self.materials_df.columns:
                # Check the actual dtype of the column
                col_dtype = self.materials_df[col].dtype
                
                # Handle different data types appropriately
                if col_dtype == 'float64' or col_dtype == 'int64':
                    # Numeric columns - convert to float, empty becomes 0.0
                    try:
                        val = float(val) if str(val).strip() else 0.0
                    except (ValueError, TypeError):
                        val = 0.0
                else:
                    # String/object columns - handle empty values
                    if val == '' or val == 'nan' or pd.isna(val):
                        val = ''
                    else:
                        val = str(val)
                
                self.materials_df.at[idx[0], col] = val
        
        # Re-check SN extraction after edit
        model = self.materials_df.at[idx[0], '모델명']
        sn = self.materials_df.at[idx[0], 'SN']
        new_model, new_sn = self.extract_sn_from_model(model, sn)
        self.materials_df.at[idx[0], '모델명'] = new_model
        self.materials_df.at[idx[0], 'SN'] = new_sn
        
        # Save to Excel
        self.save_data()
        
        # Refresh everything
        self.update_stock_view()
        self.update_registration_combos()
        self.update_material_combo()
        
        messagebox.showinfo("완료", "자재 정보가 성공적으로 수정되었습니다.")

    
    def calculate_current_stock(self, mat_id):
        """Calculate current stock for a material based on transactions"""
        # Filter transactions for this material
        mat_trans = self.transactions_df[self.transactions_df['MaterialID'] == mat_id]
        
        # Since we normalized OUT to negative in load_data, we can just sum all quantities
        net_trans_qty = mat_trans['Quantity'].sum() if not mat_trans.empty else 0
        
        # Get the current stored quantity from materials_df
        mat = self.materials_df[self.materials_df['MaterialID'] == mat_id]
        if not mat.empty:
            stored_qty = mat.iloc[0].get('수량', 0)
            # Check if stored_qty is NaN and treat it as 0
            if pd.isna(stored_qty):
                stored_qty = 0
            return stored_qty + net_trans_qty
        return 0

    def update_stock_view(self):
        # Clear current view
        for item in self.stock_tree.get_children():
            self.stock_tree.delete(item)
        
        search_term = self.search_var.get().lower() if hasattr(self, 'search_var') else ''
        
        # Helper function to safely get value and replace NaN
        def safe_get(val, default=''):
            cleaned = self.clean_nan(val)
            if cleaned == '' and str(val).lower() in ['nan', 'none', 'null', 'nan.0', '0.0', '-0.0', '']:
                return default
            return val if pd.notna(val) else default
        
        def to_f(val):
            if pd.isna(val) or val is None: return 0.0
            try:
                s = str(val).replace(',', '').strip()
                if not s: return 0.0
                return float(s)
            except:
                return 0.0
        
        # Calculate current stock
        stock_summary = []
        for _, mat in self.materials_df.iterrows():
            if mat.get('Active', True) == False:
                continue
            
            mat_id = mat['MaterialID']
            current_stock = self.calculate_current_stock(mat_id)
            
            # Note: We NO LONGER update materials_df['수량'] here to avoid infinite deduction on refresh.
            # materials_df['수량'] stays as the base/initial stock, and current_stock is calculated on-the-fly.
            
            stock_summary.append((
                mat_id,
                safe_get(mat.get('회사코드', ''), ''),
                safe_get(mat.get('관리품번', ''), ''),
                safe_get(mat.get('품목명', ''), ''),
                safe_get(mat.get('SN', ''), ''),
                safe_get(mat.get('창고', ''), ''),
                safe_get(mat.get('모델명', ''), ''),
                safe_get(mat.get('규격', ''), ''),
                safe_get(mat.get('품목군코드', ''), ''),
                safe_get(mat.get('공급업체', ''), ''),
                safe_get(mat.get('제조사', ''), ''),
                safe_get(mat.get('제조국', ''), ''),
                f"{to_f(mat.get('가격', 0)):,.0f}",
                f"{to_f(mat.get('원가', 0)):,.0f}",
                safe_get(mat.get('관리단위', 'EA'), 'EA'),
                f"{to_f(current_stock):g}",
                f"{to_f(mat.get('재고하한', 0)):g}"
            ))
        
        # Filter by search term and dropdowns
        # Filter by search term and dropdowns (stripped lowercase comparison)
        filter_co = str(self.cb_filter_co.get()).strip() if self.cb_filter_co.get() else "전체"
        filter_class = str(self.cb_filter_class.get()).strip() if self.cb_filter_class.get() else "전체"
        filter_mfr = str(self.cb_filter_mfr.get()).strip() if self.cb_filter_mfr.get() else "전체"
        filter_name = str(self.cb_filter_name.get()).strip() if self.cb_filter_name.get() else "전체"
        filter_sn = str(self.cb_filter_sn.get()).strip() if self.cb_filter_sn.get() else "전체"
        filter_model = str(self.cb_filter_model.get()).strip() if self.cb_filter_model.get() else "전체"
        filter_eq = str(self.cb_filter_eq.get()).strip() if self.cb_filter_eq.get() else "전체"
        
        for row in stock_summary:
            # Dropdown Filters
            if filter_co != "전체" and str(row[1]) != filter_co: continue
            if filter_class != "전체" and str(row[8]) != filter_class: continue
            if filter_mfr != "전체" and str(row[10]) != filter_mfr: continue
            if filter_name != "전체" and str(row[3]) != filter_name: continue
            if filter_sn != "전체" and str(row[4]) != filter_sn: continue
            if filter_model != "전체" and str(row[6]) != filter_model: continue
            if filter_eq != "전체" and str(row[2]) != filter_eq: continue
            
            # General Search Term
            if search_term:
                row_str = ' '.join(str(x).lower() for x in row)
                if search_term not in row_str:
                    continue
            self.stock_tree.insert('', tk.END, values=row)

    def setup_inout_tab(self):
        # Main PanedWindow (Vertical): Top (Registration) vs Bottom (History)
        self.inout_paned = ttk.Panedwindow(self.tab_inout, orient='vertical')
        self.inout_paned.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Horizontal PanedWindow for Registration Sections (Side-by-Side)
        self.inout_top_paned = ttk.Panedwindow(self.inout_paned, orient='horizontal')
        self.inout_paned.add(self.inout_top_paned, weight=1)
        
        # Save sash position on adjustment
        self.inout_paned.bind("<ButtonRelease-1>", lambda e: self.save_tab_config())
        
        # Side 1: New Material Registration
        reg_container = ttk.Frame(self.inout_top_paned)
        self.inout_top_paned.add(reg_container, weight=1)
        
        # Frame for Registration (Directly packed)
        reg_frame = ttk.LabelFrame(reg_container, text="자재 신규 등록")
        reg_frame.pack(fill='both', expand=True, padx=10, pady=5)
        reg_frame.grid_columnconfigure(1, weight=1)
        reg_frame.grid_columnconfigure(3, weight=1)
        
        # Row 0
        ttk.Label(reg_frame, text="회사코드:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.cb_co_code = ttk.Combobox(reg_frame, width=15)
        self.cb_co_code.grid(row=0, column=1, padx=5, pady=2, sticky='ew')
        self.cb_co_code.bind('<Return>', lambda e: self.cb_eq_code.focus_set())
        
        ttk.Label(reg_frame, text="설비코드:").grid(row=0, column=2, padx=5, pady=2, sticky='w')
        self.cb_eq_code = ttk.Combobox(reg_frame, width=20)
        self.cb_eq_code.grid(row=0, column=3, padx=5, pady=2, sticky='ew')
        self.cb_eq_code.bind('<Return>', lambda e: self.cb_item_name.focus_set())
        
        # Row 1
        ttk.Label(reg_frame, text="품목명:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.cb_item_name = ttk.Combobox(reg_frame, width=15)
        self.cb_item_name.grid(row=1, column=1, padx=5, pady=2, sticky='ew')
        self.cb_item_name.bind('<Return>', lambda e: self.ent_sn.focus_set())
        
        ttk.Label(reg_frame, text="SN번호:").grid(row=1, column=2, padx=5, pady=2, sticky='w')
        self.ent_sn = ttk.Entry(reg_frame, width=20)
        self.ent_sn.grid(row=1, column=3, padx=5, pady=2, sticky='ew')
        self.ent_sn.bind('<Return>', lambda e: self.cb_class.focus_set())
        
        # Row 2
        ttk.Label(reg_frame, text="분류:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.cb_class = ttk.Combobox(reg_frame, width=15)
        self.cb_class.grid(row=2, column=1, padx=5, pady=2, sticky='ew')
        self.cb_class.bind('<Return>', lambda e: self.cb_spec.focus_set())
        
        ttk.Label(reg_frame, text="규격:").grid(row=2, column=2, padx=5, pady=2, sticky='w')
        self.cb_spec = ttk.Combobox(reg_frame, width=20)
        self.cb_spec.grid(row=2, column=3, padx=5, pady=2, sticky='ew')
        self.cb_spec.bind('<Return>', lambda e: self.cb_unit.focus_set())
        
        # Row 3
        ttk.Label(reg_frame, text="단위:").grid(row=3, column=0, padx=5, pady=2, sticky='w')
        self.cb_unit = ttk.Combobox(reg_frame, width=15)
        self.cb_unit.grid(row=3, column=1, padx=5, pady=2, sticky='ew')
        self.cb_unit.bind('<Return>', lambda e: self.cb_supplier.focus_set())
        
        ttk.Label(reg_frame, text="공급업자:").grid(row=3, column=2, padx=5, pady=2, sticky='w')
        self.cb_supplier = ttk.Combobox(reg_frame, width=20)
        self.cb_supplier.grid(row=3, column=3, padx=5, pady=2, sticky='ew')
        self.cb_supplier.bind('<Return>', lambda e: self.cb_mfr.focus_set())
        
        # Row 4
        ttk.Label(reg_frame, text="제조사:").grid(row=4, column=0, padx=5, pady=2, sticky='w')
        self.cb_mfr = ttk.Combobox(reg_frame, width=15)
        self.cb_mfr.grid(row=4, column=1, padx=5, pady=2, sticky='ew')
        self.cb_mfr.bind('<Return>', lambda e: self.cb_origin.focus_set())
        
        ttk.Label(reg_frame, text="제조국:").grid(row=4, column=2, padx=5, pady=2, sticky='w')
        self.cb_origin = ttk.Combobox(reg_frame, width=20)
        self.cb_origin.grid(row=4, column=3, padx=5, pady=2, sticky='ew')
        self.cb_origin.bind('<Return>', lambda e: self.ent_reorder.focus_set())
        
        # Row 5
        ttk.Label(reg_frame, text="재주문 수준:").grid(row=5, column=0, padx=5, pady=2, sticky='w')
        self.ent_reorder = ttk.Entry(reg_frame, width=10)
        self.ent_reorder.grid(row=5, column=1, padx=5, pady=2, sticky='ew')
        self.ent_reorder.insert(0, "0")
        self.ent_reorder.bind('<Return>', lambda e: self.ent_init.focus_set())
        
        ttk.Label(reg_frame, text="초기재고:").grid(row=5, column=2, padx=5, pady=2, sticky='w')
        self.ent_init = ttk.Entry(reg_frame, width=15)
        self.ent_init.grid(row=5, column=3, padx=5, pady=2, sticky='ew')
        self.ent_init.insert(0, "0")
        self.ent_init.bind('<Return>', lambda e: self.ent_price.focus_set())
        
        ttk.Label(reg_frame, text="단가(가격):").grid(row=6, column=0, padx=5, pady=2, sticky='w')
        self.ent_price = ttk.Entry(reg_frame, width=15)
        self.ent_price.grid(row=6, column=1, padx=5, pady=2, sticky='ew')
        self.ent_price.insert(0, "0")
        self.ent_price.bind('<Return>', lambda e: self.ent_cost.focus_set())
        
        ttk.Label(reg_frame, text="원가:").grid(row=6, column=2, padx=5, pady=2, sticky='w')
        self.ent_cost = ttk.Entry(reg_frame, width=15)
        self.ent_cost.grid(row=6, column=3, padx=5, pady=2, sticky='ew')
        self.ent_cost.insert(0, "0")
        self.ent_cost.bind('<Return>', lambda e: btn_reg.focus_set())
        
        btn_reg = ttk.Button(reg_frame, text="자재 등록", command=self.register_material)
        btn_reg.grid(row=7, column=0, columnspan=4, pady=10)
        
        # Side 2: Transaction Entry
        trans_reg_container = ttk.Frame(self.inout_top_paned)
        self.inout_top_paned.add(trans_reg_container, weight=1)
        
        # Frame for In/Out Transaction (Directly packed)
        trans_frame = ttk.LabelFrame(trans_reg_container, text="입출고 기록")
        trans_frame.pack(fill='both', expand=True, padx=10, pady=5)
        trans_frame.grid_columnconfigure(1, weight=1)
        trans_frame.grid_columnconfigure(3, weight=1)

        # Bottom frame for history (Occupies full width at bottom)
        history_container = ttk.Frame(self.inout_paned)
        self.inout_paned.add(history_container, weight=2)
        
        ttk.Label(trans_frame, text="자재 선택:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.cb_material = ttk.Combobox(trans_frame, width=45, font=('Malgun Gothic', 10))
        self.cb_material.grid(row=0, column=1, padx=5, pady=2, columnspan=3, sticky='ew')
        self.cb_material.bind('<<ComboboxSelected>>', self.on_material_selected)
        self.cb_material.bind('<Return>', lambda e: self.cb_type.focus_set())
        
        # Register autocomplete for material selection
        self.enable_autocomplete(self.cb_material, values_list=[]) # Initial empty, will be updated by update_material_combo
        self.update_material_combo()
        
        # Model List display - Moved to row 4 for narrower fit
        ttk.Label(trans_frame, text="관련 모델명:").grid(row=4, column=0, padx=5, pady=2, sticky='nw')
        self.list_models = tk.Listbox(trans_frame, height=3, width=42, font=('Malgun Gothic', 9))
        self.list_models.grid(row=4, column=1, columnspan=3, padx=5, pady=2, sticky='ew')
        
        # Add scrollbar for model list
        model_vsb = ttk.Scrollbar(trans_frame, orient="vertical", command=self.list_models.yview)
        model_vsb.grid(row=4, column=4, sticky='ns', pady=2)
        self.list_models.config(yscrollcommand=model_vsb.set)
        
        ttk.Label(trans_frame, text="구분:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.cb_type = ttk.Combobox(trans_frame, values=["IN", "OUT"], state="readonly", width=15)
        self.cb_type.grid(row=1, column=1, padx=5, pady=2, sticky='ew')
        self.cb_type.set("OUT")
        self.cb_type.bind('<Return>', lambda e: self.ent_qty.focus_set())
        
        ttk.Label(trans_frame, text="수량:").grid(row=1, column=2, padx=5, pady=2, sticky='w')
        self.ent_qty = ttk.Entry(trans_frame, width=30)
        self.ent_qty.grid(row=1, column=3, padx=5, pady=2, sticky='ew')
        self.ent_qty.bind('<Return>', lambda e: self.cb_trans_site.focus_set())
        
        ttk.Label(trans_frame, text="현장:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.cb_trans_site = ttk.Combobox(trans_frame, width=28, values=self.sites)
        self.cb_trans_site.grid(row=2, column=1, padx=5, pady=2, sticky='ew')
        self.cb_trans_site.bind('<FocusOut>', lambda e: self.auto_save_to_list(e, self.cb_trans_site, self.sites, 'sites'))
        self.cb_trans_site.bind('<Return>', lambda e: self._on_trans_site_return(e))
        
        ttk.Label(trans_frame, text="창고:").grid(row=2, column=2, padx=5, pady=2, sticky='w')
        self.cb_warehouse = ttk.Combobox(trans_frame, width=28, values=self.warehouses)
        self.cb_warehouse.grid(row=2, column=3, padx=5, pady=2, sticky='ew')
        self.cb_warehouse.bind('<FocusOut>', lambda e: self.auto_save_to_list(e, self.cb_warehouse, self.warehouses, 'warehouses'))
        self.cb_warehouse.bind('<Return>', lambda e: self._on_warehouse_return(e))
        
        ttk.Label(trans_frame, text="담당자:").grid(row=3, column=0, padx=5, pady=2, sticky='w')
        self.ent_user = ttk.Combobox(trans_frame, width=28, values=getattr(self, 'users', []))
        self.ent_user.grid(row=3, column=1, padx=5, pady=2, sticky='ew')
        self.ent_user.bind('<FocusOut>', lambda e: self.auto_save_to_list(e, self.ent_user, self.users, 'users'))
        self.ent_user.bind('<Return>', lambda e: self._on_user_return(e))
        
        ttk.Label(trans_frame, text="비고:").grid(row=3, column=2, padx=5, pady=2, sticky='w')
        self.ent_note = ttk.Entry(trans_frame, width=30)
        self.ent_note.grid(row=3, column=3, padx=5, pady=2, sticky='ew')
        self.ent_note.bind('<Return>', lambda e: btn_trans.focus_set())
        
        btn_trans = ttk.Button(trans_frame, text="기록 저장", command=self.add_transaction)
        btn_trans.grid(row=5, column=0, columnspan=4, pady=10)
        
        # Frame for displaying transaction history
        history_frame = ttk.LabelFrame(history_container, text="최근 입출고 내역")
        history_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Filter/Control frame for history
        history_ctrl_frame = ttk.Frame(history_frame)
        history_ctrl_frame.pack(fill='x', padx=5, pady=2)
        
        # First row of filters
        filter_row1 = ttk.Frame(history_ctrl_frame)
        filter_row1.pack(fill='x', pady=2)
        
        btn_del_trans = ttk.Button(filter_row1, text="선택 항목 삭제", command=self.delete_transaction_entry)
        btn_del_trans.pack(side='left', padx=5)
        
        btn_export_trans = ttk.Button(filter_row1, text="엑셀 내보내기", command=self.export_transaction_history)
        btn_export_trans.pack(side='left', padx=5)
        
        # Second row of filters
        filter_row2 = ttk.Frame(history_ctrl_frame)
        filter_row2.pack(fill='x', pady=2)
        
        ttk.Label(filter_row2, text="품목명 필터:").pack(side='left', padx=5)
        self.cb_trans_filter_mat = ttk.Combobox(filter_row2, width=35, state="readonly")
        self.cb_trans_filter_mat.pack(side='left', padx=5)
        self.cb_trans_filter_mat.bind('<<ComboboxSelected>>', lambda e: self.update_transaction_view())
        
        ttk.Label(filter_row2, text="현장 필터:").pack(side='left', padx=(20, 5))
        self.cb_trans_filter_site = ttk.Combobox(filter_row2, width=12, state="readonly")
        self.cb_trans_filter_site.pack(side='left', padx=5)
        self.cb_trans_filter_site.bind('<<ComboboxSelected>>', lambda e: self.update_transaction_view())
        
        # Third row of filters
        filter_row3 = ttk.Frame(history_ctrl_frame)
        filter_row3.pack(fill='x', pady=2)
        
        ttk.Label(filter_row3, text="차량번호 필터:").pack(side='left', padx=5)
        self.cb_trans_filter_vehicle = ttk.Combobox(filter_row3, width=12, state="readonly")
        self.cb_trans_filter_vehicle.pack(side='left', padx=5)
        self.cb_trans_filter_vehicle.bind('<<ComboboxSelected>>', lambda e: self.update_transaction_view())
        
        # Treeview for history
        tree_scroll_frame = ttk.Frame(history_frame)
        tree_scroll_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        inout_vsb = ttk.Scrollbar(tree_scroll_frame, orient="vertical")
        inout_hsb = ttk.Scrollbar(tree_scroll_frame, orient="horizontal")
        
        columns = ('날짜', '현장', '품목명', '모델명', '구분', '수량', '재고', '담당자', '차량번호', '주행거리', '차량점검', '차량비고', '비고')
        self.inout_tree = ttk.Treeview(tree_scroll_frame, columns=columns, show='headings', height=10,
                                       yscrollcommand=inout_vsb.set, xscrollcommand=inout_hsb.set)
        
        inout_vsb.config(command=self.inout_tree.yview)
        inout_hsb.config(command=self.inout_tree.xview)
        
        col_widths = [150, 120, 180, 150, 70, 70, 70, 100, 120, 100, 200, 150, 200]
        for col, width in zip(columns, col_widths):
            self.inout_tree.heading(col, text=col)
            self.inout_tree.column(col, width=width, minwidth=50, stretch=True, anchor='center')
        
        self.inout_tree.grid(row=0, column=0, sticky='nsew')
        inout_vsb.grid(row=0, column=1, sticky='ns')
        inout_hsb.grid(row=1, column=0, sticky='ew')
        
        tree_scroll_frame.grid_rowconfigure(0, weight=1)
        tree_scroll_frame.grid_columnconfigure(0, weight=1)
        
        # Initial populate
        self.update_transaction_view()
        
        # Set initial sash position for better balance
        self.inout_paned.after(200, self._ensure_inout_sash_visibility)

    def _ensure_inout_sash_visibility(self):
        """Ensure the inout sash position is properly set"""
        try:
            if hasattr(self, 'inout_paned'):
                self.inout_paned.update_idletasks()
                total_h = self.inout_paned.winfo_height()
                if total_h > 200:
                    # Set vertical sash to 40% of total height (registration area smaller)
                    new_pos = int(total_h * 0.4)
                    self.inout_paned.sashpos(0, new_pos)
                    
                    # Set horizontal sash for side-by-side
                    total_w = self.inout_top_paned.winfo_width()
                    if total_w > 200:
                        self.inout_top_paned.sashpos(0, int(total_w * 0.45))
                    print(f"Set inout sash to {new_pos} (total height: {total_h})")
        except Exception as e:
            print(f"Error ensuring inout sash visibility: {e}")


    def get_material_display_name(self, mat_id):
        """Get formatted material name as '품목명 (SN: SN번호) - 규격'"""
        if self.materials_df.empty:
            return f"ID: {mat_id}"
            
        mat_row = self.materials_df[self.materials_df['MaterialID'] == mat_id]
        if mat_row.empty:
            # Try numeric fallback just in case some IDs are still strings
            try:
                num_id = int(float(mat_id))
                mat_row = self.materials_df[self.materials_df['MaterialID'] == num_id]
            except (ValueError, TypeError):
                pass
                
        if mat_row.empty:
            return f"ID: {mat_id}"
            
        mat = mat_row.iloc[0]
        name = mat['품목명']
        sn = mat.get('SN', '')
        spec = mat.get('규격', '')
        
        # Build display string
        display = name
        
        # Add SN if exists
        if sn and pd.notna(sn) and str(sn).strip():
            display += f" (SN: {sn})"
        
        # Add specification if exists
        if spec and pd.notna(spec) and str(spec).strip():
            display += f" - {spec}"
            
        return display

    def update_material_combo(self):
        """Update both In/Out and Daily Usage material comboboxes with unified list"""
        mat_list = []
        if not self.materials_df.empty:
            # Create list with unified display format
            for _, mat in self.materials_df.iterrows():
                if mat.get('Active', True) == False:
                    continue
                display = self.get_material_display_name(mat['MaterialID'])
                mat_list.append(display)
        
        # [NEW] Also collect historical material names from daily usage history
        history_materials = []
        if not self.daily_usage_df.empty and 'MaterialID' in self.daily_usage_df.columns:
            unique_mat_ids = self.daily_usage_df['MaterialID'].dropna().unique()
            for mat_id in unique_mat_ids:
                name = self.get_material_display_name(mat_id)
                if name: history_materials.append(name)
        
        # Merge database items with centralized films list and historical materials, unique and sort
        all_vals = list(set([str(m) for m in mat_list + self.carestream_films + history_materials if pd.notna(m) and str(m).strip()]))
        all_vals.sort()
        
        # Update ComboBoxes
        if hasattr(self, 'cb_material'):
            self.cb_material['values'] = all_vals
        if hasattr(self, 'cb_daily_material'):
            self.cb_daily_material['values'] = all_vals
            
        if hasattr(self, 'cb_daily_equip'):
            # [NEW] Collect history of equipment names from daily usage records
            history_equip = []
            if not self.daily_usage_df.empty and '장비명' in self.daily_usage_df.columns:
                history_equip = [str(x).strip() for x in self.daily_usage_df['장비명'].dropna().unique().tolist() if str(x).strip()]
                
            # Combine custom equipments, history, and the full material display list (all_vals)
            combined_equip = list(set([str(e).strip() for e in self.equipments + history_equip + all_vals if pd.notna(e) and str(e).strip()]))
            combined_equip.sort()
            self.cb_daily_equip['values'] = combined_equip
            self.equipment_suggestions = combined_equip
        
        self.materials_display_list = all_vals


        if hasattr(self, 'cb_trans_filter_mat'):
            self.cb_trans_filter_mat['values'] = ["전체"] + all_vals
            if not self.cb_trans_filter_mat.get():
                self.cb_trans_filter_mat.set("전체")
        
        if hasattr(self, 'cb_trans_filter_site'):
            self.cb_trans_filter_site['values'] = ["전체"] + sorted(self.sites)
            if not self.cb_trans_filter_site.get():
                self.cb_trans_filter_site.set("전체")
        
        if hasattr(self, 'cb_trans_filter_vehicle'):
            # Get unique vehicle numbers from transactions_df
            if '차량번호' in self.transactions_df.columns:
                vehicle_numbers = self.transactions_df['차량번호'].dropna().astype(str).str.strip().unique()
                vehicle_list = sorted([v for v in vehicle_numbers if v and v != 'nan'])
            else:
                vehicle_list = []
            
            # Add sample vehicle numbers for testing if no data exists
            if not vehicle_list:
                vehicle_list = ['12가1234', '34나5678', '89다9012']
            
            self.cb_trans_filter_vehicle['values'] = ["전체"] + vehicle_list
            if not self.cb_trans_filter_vehicle.get():
                self.cb_trans_filter_vehicle.set("전체")

    def update_registration_combos(self):
        """Update registration comboboxes with unique values from database and centralized list"""
        # 1. Update registration fields from database
        fields = {
            '회사코드': self.cb_co_code,
            '관리품번': self.cb_eq_code,
            '품목명': self.cb_item_name,
            '품목군코드': self.cb_class,
            '규격': self.cb_spec,
            '관리단위': self.cb_unit,
            '제조사': self.cb_mfr,
            '제조국': self.cb_origin
        }
        
        # Store lists for autocomplete (Main Catalog + History)
        self.co_code_list = []
        self.eq_code_list = []
        self.item_name_list = []
        self.class_list = []
        self.spec_list = []
        self.unit_list = []
        self.mfr_list = []
        self.origin_list = []
        self.sn_list = []
        self.model_list = []

        attr_mapping = {
            '회사코드': 'co_code_list',
            '관리품번': 'eq_code_list',
            '품목명': 'item_name_list',
            '품목군코드': 'class_list',
            '규격': 'spec_list',
            '관리단위': 'unit_list',
            '제조사': 'mfr_list',
            '제조국': 'origin_list',
            'SN': 'sn_list',
            '모델명': 'model_list'
        }

        for col, list_attr in attr_mapping.items():
            vals = []
            if not self.materials_df.empty and col in self.materials_df.columns:
                active_df = self.materials_df[self.materials_df.get('Active', True) != False]
                unique_vals = active_df[col].dropna().unique()
                vals = sorted([str(v).strip() for v in unique_vals if v and str(v).strip()])
            
            # Update instance attributes for autocomplete
            setattr(self, list_attr, vals)
            
            # 1. Update registration fields (without "전체")
            if col in fields:
                fields[col]['values'] = vals

            # 2. Update Stock View filters (with "전체")
            filter_map = {
                '회사코드': getattr(self, 'cb_filter_co', None),
                '품목군코드': getattr(self, 'cb_filter_class', None),
                '제조사': getattr(self, 'cb_filter_mfr', None),
                '품목명': getattr(self, 'cb_filter_name', None),
                'SN': getattr(self, 'cb_filter_sn', None),
                '모델명': getattr(self, 'cb_filter_model', None),
                '관리품번': getattr(self, 'cb_filter_eq', None)
            }
            if col in filter_map and filter_map[col] is not None:
                combo = filter_map[col]
                combo['values'] = ["전체"] + vals
                if not combo.get():
                    combo.set("전체")
        
        # Add Carestream film options to 품목명 (Registration AND Stock Filter)
        if hasattr(self, 'cb_item_name') or hasattr(self, 'cb_filter_name'):
            all_mat_vals = sorted(list(set([str(mat) for mat in self.item_name_list + self.carestream_films if pd.notna(mat) and str(mat).strip()])))
            self.item_name_list[:] = all_mat_vals
            
            if hasattr(self, 'cb_item_name'):
                self.cb_item_name['values'] = all_mat_vals
            
            if hasattr(self, 'cb_filter_name'):
                self.cb_filter_name['values'] = ["전체"] + all_mat_vals
                if not self.cb_filter_name.get():
                    self.cb_filter_name.set("전체")
        
        # 3. Final view update now that filters are set to "전체"
        import threading
        self.root.after(100, self.update_stock_view)

    def register_material(self):
        co_code = self.cb_co_code.get()
        item_name = self.cb_item_name.get()
        eq_code = self.cb_eq_code.get()
        sn = self.ent_sn.get()
        classification = self.cb_class.get()
        spec = self.cb_spec.get()
        unit = self.cb_unit.get()
        supplier = self.cb_supplier.get()
        manufacturer = self.cb_mfr.get()
        origin = self.cb_origin.get()
        
        try:
            init_stock = float(self.ent_init.get())
            reorder_point = float(self.ent_reorder.get())
            price_val = float(self.ent_price.get() if self.ent_price.get() else 0)
            cost_val = float(self.ent_cost.get() if self.ent_cost.get() else 0)
        except ValueError:
            messagebox.showwarning("입력 오류", "재고, 단가, 원가 정보는 숫자여야 합니다.")
            return
            
        if not item_name:
            messagebox.showwarning("입력 오류", "품목명을 입력해주세요.")
            return
        
        # Generate MaterialID
        if self.materials_df.empty:
            mat_id = 1
        else:
            mat_id = self.materials_df['MaterialID'].max() + 1
        
        # Extract SN from Model Name if present
        model_name = '' # Default empty
        new_model, new_sn = self.extract_sn_from_model(model_name, sn)
        # Note: In register_material, model_name is currently not an input field, but SN is.
        # If the user adds model name to registration later, this will handle it.
        
        new_row = {
            'MaterialID': mat_id,
            '회사코드': co_code,
            '관리품번': eq_code,
            '품목명': item_name,
            'SN': new_sn,
            '창고': '',
            '모델명': new_model,
            '규격': spec,
            '품목군코드': classification,
            '공급업체': supplier,
            '제조사': manufacturer,
            '제조국': origin,
            '가격': price_val,
            '원가': cost_val,
            '관리단위': unit if unit else 'EA',
            '수량': init_stock,
            '재고하한': reorder_point
        }
        
        self.materials_df = pd.concat([self.materials_df, pd.DataFrame([new_row])], ignore_index=True)
        self.save_data()
        self.update_material_combo()
        self.update_registration_combos()
        self.update_stock_view()
        messagebox.showinfo("완료", f"'{item_name}' 자재가 등록되었습니다.")
        
        # Clear entries
        self.cb_co_code.set('')
        self.cb_eq_code.set('')
        self.cb_item_name.set('')
        self.ent_sn.delete(0, tk.END)
        self.cb_class.set('')
        self.cb_spec.set('')
        self.cb_unit.set('')
        self.cb_supplier.set('')
        self.cb_mfr.set('')
        self.cb_origin.set('')
        self.ent_reorder.delete(0, tk.END)
        self.ent_reorder.insert(0, "0")
        self.ent_init.delete(0, tk.END)
        self.ent_init.insert(0, "0")

    def add_transaction(self):
        """Record an IN or OUT transaction"""
        try:
            mat_selection = self.cb_material.get().strip()
            t_type = self.cb_type.get()
            user = self.ent_user.get().strip()
            
            if not mat_selection:
                messagebox.showwarning("입력 오류", "자재를 선택해주세요.")
                return
            if not t_type:
                messagebox.showwarning("입력 오류", "구분(입고/출고)을 선택해주세요.")
                return
                
            try:
                qty_str = self.ent_qty.get().strip()
                if not qty_str:
                    messagebox.showwarning("입력 오류", "수량을 입력해주세요.")
                    return
                qty = float(qty_str)
            except ValueError:
                messagebox.showwarning("입력 오류", "수량은 숫자여야 합니다.")
                return
                
            note = self.ent_note.get().strip()
            
            # Extract pure material name from selection
            mat_name = mat_selection
            if " - " in mat_name: mat_name = mat_name.split(" - ")[0]
            if " (SN: " in mat_name: mat_name = mat_name.split(" (SN: ")[0]
            pure_mat_name = mat_name.strip()
            
            # Find MaterialID
            # Ensure MaterialID is treated consistently
            mat_rows = self.materials_df[self.materials_df['품목명'] == pure_mat_name]
            
            if mat_rows.empty:
                # If exact match fails, try a case-insensitive or stripped match
                mat_rows = self.materials_df[self.materials_df['품목명'].str.strip() == pure_mat_name]
            
            if mat_rows.empty:
                messagebox.showerror("오류", f"'{pure_mat_name}' 자재를 찾을 수 없습니다.\n먼저 자재를 등록하거나 정확한 명칭을 입력해주세요.")
                return
                
            mat_id = mat_rows['MaterialID'].values[0]
            
            # Update Warehouse in materials_df
            warehouse = str(self.cb_warehouse.get()).strip()
            if warehouse:
                # Type-safe assignment
                if '창고' in self.materials_df.columns:
                    mask = self.materials_df['MaterialID'] == mat_id
                    if mask.any():
                        self.materials_df.loc[mask, '창고'] = warehouse
            
            # Create transaction record
            new_trans = {
                'Date': datetime.datetime.now(),
                'MaterialID': mat_id,
                'Type': t_type,
                'Quantity': qty if t_type == 'IN' else -qty,
                'Note': note,
                'User': user,
                'Site': self.cb_trans_site.get() if hasattr(self, 'cb_trans_site') else '',
                '차량번호': '',
                '주행거리': '',
                '차량점검': '',
                '차량비고': ''
            }
            
            # Add to dataframe
            self.transactions_df = pd.concat([self.transactions_df, pd.DataFrame([new_trans])], ignore_index=True)
            
            # Save all data (force sync)
            self.save_data()
            
            # Check if it was really added (for debug feedback)
            last_count = len(self.transactions_df)
            
            # Refresh views
            self.update_stock_view()
            self.update_transaction_view()
            self.update_material_combo() 
            
            # Auto-save site/user to lists
            site_value = new_trans['Site']
            if site_value and site_value not in self.sites:
                self.sites.append(site_value)
                self.sites.sort()
                self.save_tab_config()
            
            if user and user not in self.users:
                self.users.append(user)
                self.users.sort()
                self.save_tab_config()
            
            # Success feedback
            messagebox.showinfo("완료", f"{pure_mat_name} {t_type} 처리되었습니다.\n(전체 기록 수: {last_count}개)")
            
            # Clear UI fields
            self.ent_qty.delete(0, tk.END)
            self.ent_note.delete(0, tk.END)
            if hasattr(self, 'ent_user') and hasattr(self.ent_user, 'set'):
                self.ent_user.set('')
            elif hasattr(self, 'ent_user'):
                self.ent_user.delete(0, tk.END)
            
            if hasattr(self, 'cb_warehouse'): self.cb_warehouse.set('')
            if hasattr(self, 'cb_trans_site'): self.cb_trans_site.set('')
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            self.show_error_dialog("저장 오류", f"기록 저장 중 기술적인 오류가 발생했습니다:\n{e}\n\n상세 정보:\n{error_details}")
        
        # Ensure view is updated regardless
        self.update_transaction_view()

    def update_transaction_view(self):
        """Populate the transaction history Treeview"""
        try:
            if not hasattr(self, 'inout_tree'):
                return
                
            # Clear current view
            for item in self.inout_tree.get_children():
                self.inout_tree.delete(item)
                
            if self.transactions_df.empty:
                return
                
            # Update frame title with count
            if hasattr(self, 'scrollable_frame'):
                # Try to find the history LabelFrame to update its text
                for child in self.scrollable_frame.winfo_children():
                    if isinstance(child, ttk.LabelFrame) and child.cget("text").startswith("최근 입출고 내역"):
                        child.configure(text=f"최근 입출고 내역 (총 {len(self.transactions_df)}건)")
            
            # Display last 500 transactions, descending by date
            df_to_show = self.transactions_df.copy()
            
            # 1. Pre-calculate Running Balances for ALL transactions to show historical accuracy
            initial_stocks = self.materials_df.set_index('MaterialID')['수량'].fillna(0).to_dict()
            df_running = df_to_show.copy().sort_values(by='Date')
            
            # Map IN/OUT to signed values
            # Since we already normalized OUT to negative in load_data and save logic, 
            # we can use Quantity directly.
            df_running['SignedQty'] = df_running['Quantity']
            
            # Calculate cumulative sum per material
            df_running['HistoricalStock'] = df_running.groupby('MaterialID')['SignedQty'].cumsum()
            
            # Add initial stock offset
            df_running['HistoricalStock'] = df_running.apply(lambda r: initial_stocks.get(r['MaterialID'], 0) + r['HistoricalStock'], axis=1)
            
            # Merge back to the copy we want to filter/show
            df_to_show = df_to_show.merge(df_running[['HistoricalStock']], left_index=True, right_index=True)
            
            # Apply material filter if selected (Optimized)
            if hasattr(self, 'cb_trans_filter_mat'):
                selected_mat = self.cb_trans_filter_mat.get()
                if selected_mat and selected_mat != "전체":
                    # Pre-calculate matching IDs to avoid repetitive lookups
                    matching_ids = []
                    for mat_id in df_to_show['MaterialID'].unique():
                        if pd.isna(mat_id): continue
                        if self.get_material_display_name(mat_id) == selected_mat:
                            matching_ids.append(mat_id)
                    df_to_show = df_to_show[df_to_show['MaterialID'].isin(matching_ids)]
            
            # Apply site filter if selected
            if hasattr(self, 'cb_trans_filter_site'):
                selected_site = self.cb_trans_filter_site.get()
                if selected_site and selected_site != "전체":
                    df_to_show = df_to_show[df_to_show['Site'] == selected_site]
            
            # Apply vehicle filter if selected
            if hasattr(self, 'cb_trans_filter_vehicle'):
                selected_vehicle = self.cb_trans_filter_vehicle.get()
                if selected_vehicle and selected_vehicle != "전체":
                    df_to_show = df_to_show[df_to_show['차량번호'].astype(str).str.contains(selected_vehicle, na=False, case=False, regex=False)]

            df_sorted = df_to_show.sort_values(by='Date', ascending=False, na_position='last').head(500)
            
            for idx, row in df_sorted.iterrows():
                mat_id = row['MaterialID']
                mat_name = self.get_material_display_name(mat_id)
                
                # Safe date formatting
                raw_date = row.get('Date')
                if pd.isna(raw_date):
                    date_str = "Unknown"
                elif hasattr(raw_date, 'strftime'):
                    date_str = raw_date.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    date_str = str(raw_date)
                
                # Use pre-calculated historical stock for balance view
                historical_stock = row.get('HistoricalStock', 0)
                
                # Find model name from materials_df
                model_name = ""
                mat_row = self.materials_df[self.materials_df['MaterialID'] == mat_id]
                if not mat_row.empty:
                    model_name = str(mat_row.iloc[0].get('모델명', '')).replace('nan', '').replace('None', '')
                
                self.inout_tree.insert('', tk.END, values=(
                    date_str,
                    self.clean_nan(row.get('Site', '')),
                    mat_name,
                    model_name,
                    row.get('Type', ''),
                    row.get('Quantity', 0),
                    historical_stock,
                    self.clean_nan(row.get('User', '')),
                    self.clean_nan(row.get('차량번호', '')),
                    self.clean_nan(row.get('주행거리', '')),
                    self.clean_nan(row.get('차량점검', '')),
                    self.clean_nan(row.get('차량비고', '')),
                    self.clean_nan(row.get('Note', ''))
                ), tags=(str(idx),))
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Transaction View Error: {e}\n{error_details}")

    def delete_transaction_entry(self):
        """Delete selected transaction from history and refresh"""
        selection = self.inout_tree.selection()
        if not selection:
            messagebox.showwarning("선택 오류", "삭제할 기록을 선택해주세요.")
            return
            
        if not messagebox.askyesno("삭제 확인", "선택한 기록을 삭제하시겠습니까?\n(삭제 시 재고 계산에 즉시 반영됩니다.)"):
            return
            
        for item in selection:
            tags = self.inout_tree.item(item, 'tags')
            if tags:
                idx = int(tags[0])
                if idx in self.transactions_df.index:
                    self.transactions_df = self.transactions_df.drop(idx)
        
        self.save_data()
        self.update_transaction_view()
        self.update_stock_view()
        messagebox.showinfo("완료", "거래 기록이 삭제되었습니다.")

    def on_material_selected(self, event=None):
        """Update the model listbox based on selected material"""
        selection = self.cb_material.get()
        if not selection:
            return
            
        # Clear existing models
        self.list_models.delete(0, tk.END)
        
        # Extract pure material name
        mat_name = selection
        if " - " in mat_name:
            mat_name = mat_name.split(" - ")[0]
        if " (SN: " in mat_name:
            mat_name = mat_name.split(" (SN: ")[0]
        
        pure_mat_name = mat_name
        
        # Find unique models for this material in materials_df
        if not self.materials_df.empty:
            relevant_mats = self.materials_df[self.materials_df['품목명'] == pure_mat_name]
            if not relevant_mats.empty:
                unique_models = relevant_mats['모델명'].dropna().unique()
                unique_models = sorted([str(m).strip() for m in unique_models if str(m).strip()])
                
                for model in unique_models:
                    self.list_models.insert(tk.END, model)
            # If no models found, add a placeholder
                if not unique_models:
                    self.list_models.insert(tk.END, "(등록된 모델명 없음)")

    def _on_trans_site_return(self, event):
        self.auto_save_to_list(event, self.cb_trans_site, self.sites, 'sites')
        self.cb_warehouse.focus_set()

    def _on_warehouse_return(self, event):
        self.auto_save_to_list(event, self.cb_warehouse, self.warehouses, 'warehouses')
        self.ent_user.focus_set()

    def _on_user_return(self, event):
        self.auto_save_to_list(event, self.ent_user, self.users, 'users')
        self.ent_note.focus_set()


    def format_worker_summary(self, workers):
        """Format a list of workers (or a joined string) into a compact summary string with a dropdown cue"""
        if not workers: return ""
        if isinstance(workers, str):
            if " | " in workers:
                names = [w.strip() for w in workers.split(" | ") if w.strip()]
            else:
                names = [w.strip() for w in workers.split(",") if w.strip()]
        else:
            names = [self.clean_nan(w) for w in workers if self.clean_nan(w)]
            
        unique_names = sorted(list(set(names)))
        if not unique_names: return ""
        
        return f"{unique_names[0]} [▼]"

    def show_worker_popup(self, event, tree):
        """Show a floating window (dropdown-style) with full record details for the clicked row"""
        item_id = tree.identify_row(event.y)
        column_id = tree.identify_column(event.x)
        if not item_id or not column_id: return
            
        col_name = tree.heading(column_id)['text']
        if col_name != '작업자': return
            
        tags = tree.item(item_id, 'tags')
        if not tags or 'total' in tags: return
            
        try:
            full_info = []
            if tree == getattr(self, 'daily_usage_tree', None):
                try:
                    df_idx = int(tags[0])
                    if df_idx not in self.daily_usage_df.index: return
                    entry = self.daily_usage_df.loc[df_idx]
                    
                    # 1. Basic Info
                    usage_date = self._safe_format_datetime(entry.get('Date', ''), '%Y-%m-%d')
                    full_info.append(f"[기본 기록 정보]")
                    full_info.append(f"• 날짜: {usage_date}")
                    full_info.append(f"• 현장: {self.clean_nan(entry.get('Site', ''))}")
                    
                    equip = self.clean_nan(entry.get('장비명', ''))
                    if equip: full_info.append(f"• 장비명: {equip}")
                    
                    method = self.clean_nan(entry.get('검사방법', ''))
                    if method: full_info.append(f"• 검사방법: {method}")
                    
                    amount = entry.get('검사량', 0)
                    if amount and float(amount) > 0: full_info.append(f"• 검사량: {amount}")
                    
                    f_count = entry.get('FilmCount', 0)
                    if f_count and float(f_count) > 0: full_info.append(f"• RTK: {f_count}")
                    
                    full_info.append(f"• 품목명: {self.get_material_display_name(entry.get('MaterialID'))}")
                    
                    # 2. Worker Info (All 10 slots)
                    full_info.append(f"\n[작업자 및 시간 정보]")
                    for i in range(1, 11):
                        u = self.clean_nan(entry.get('User' if i==1 else f'User{i}', ''))
                        if u:
                            t = self.clean_nan(entry.get('WorkTime' if i==1 else f'WorkTime{i}', ''))
                            o = self.clean_nan(entry.get('OT' if i==1 else f'OT{i}', ''))
                            full_info.append(f"• 작업자{i}: {u} | {t} | OT: {o}")
                    
                    # 3. Cost Info (Hide zeros)
                    cost_info = []
                    for k, label in [('단가', '단가'), ('출장비', '출장비'), ('일식', '일식'), ('검사비', '검사비')]:
                        raw_val = entry.get(k, 0)
                        try:
                            if pd.isna(raw_val) or str(raw_val).lower() == 'nan': continue
                            val_float = float(str(raw_val).replace(',', ''))
                            if val_float > 0:
                                cost_info.append(f"• {label}: {int(val_float):,}")
                        except: pass
                    if cost_info:
                        full_info.append(f"\n[비용 상세]")
                        full_info.extend(cost_info)
                    
                    # 4. RTK/NDT (Hide zeros)
                    rtk_found = []
                    rtk_cats = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타"]
                    for cat in rtk_cats:
                        val = entry.get(f'RTK_{cat}', 0)
                        if val and float(val) > 0:
                            rtk_found.append(f"  - {cat}: {val}")
                    if rtk_found:
                        full_info.append(f"\n[RT 결함 상세]")
                        full_info.extend(rtk_found)
                        
                    ndt_found = []
                    ndt_mats = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
                    for mat in ndt_mats:
                        val = entry.get(f'NDT_{mat}', 0)
                        if pd.isna(val): # Fallback
                            if mat == "흑색자분": val = entry.get('NDT_자분', 0)
                            elif mat == "백색페인트": val = entry.get('NDT_페인트', 0)
                        if val and float(val) > 0:
                            ndt_found.append(f"  - {mat}: {val}")
                    if ndt_found:
                        full_info.append(f"\n[NDT 자재 사용량]")
                        full_info.extend(ndt_found)
                    
                    note = entry.get('Note', '')
                    if note and str(note).strip(): full_info.append(f"\n• 비고: {note}")
                    
                    full_info.append(f"• 입력시간: {entry.get('EntryTime', '')}")
                    
                except Exception as e: 
                    print(f"Popup error: {e}")
                    return
            else:
                # Fallback for other trees (Monthly usage, etc.)
                vals = tree.item(item_id, 'values')
                cols = tree['columns']
                try:
                    w_idx = cols.index('작업자')
                    t_idx = cols.index('작업시간')
                    raw_names = vals[w_idx]
                    raw_times = vals[t_idx]
                    full_info.append(f"작업자: {raw_names}")
                    full_info.append(f"작업시간: {raw_times}")
                except: return

            if not full_info: return
            
            # Get cell bounding box for precise positioning
            bbox = tree.bbox(item_id, column_id)
            if not bbox: return
            
            x, y, w, h = bbox
            root_x = tree.winfo_rootx() + x
            root_y = tree.winfo_rooty() + y
            
            popup = tk.Toplevel(self.root)
            popup.overrideredirect(True)
            popup.attributes("-topmost", True)
            popup.configure(bg='white', highlightbackground="#0078d7", highlightthickness=2)
            
            popup_width = max(w, 550)
            popup_height = 450
            
            # 1. Horizontal positioning (Center relative to cell)
            final_x = root_x - (popup_width - w) // 2
            
            # 2. Vertical positioning (Default: Below the cell)
            final_y = root_y + h + 1
            
            # Simple boundary check removed for multi-monitor support
            # Initial relative offsets from the cell root
            popup._rel_offset_x = final_x - root_x
            popup._rel_offset_y = final_y - root_y

            popup.geometry(f"{popup_width}x{popup_height}+{int(final_x)}+{int(final_y)}")
            
            # --- Custom Dragging Support ---
            def start_drag(event):
                popup._drag_data = {"x": event.x, "y": event.y}
            
            def do_drag(event):
                if hasattr(popup, '_drag_data'):
                    dx = event.x - popup._drag_data["x"]
                    dy = event.y - popup._drag_data["y"]
                    new_x = popup.winfo_x() + dx
                    new_y = popup.winfo_y() + dy
                    popup.geometry(f"+{new_x}+{new_y}")
                    
                    # Update relative offsets after manual drag so sync continues from new spot
                    try:
                        curr_bbox = tree.bbox(item_id, column_id)
                        if curr_bbox:
                            cx, cy, cw, ch = curr_bbox
                            # Important: Update relative to the CURRENT root position
                            popup._rel_offset_x = new_x - (tree.winfo_rootx() + cx)
                            popup._rel_offset_y = new_y - (tree.winfo_rooty() + cy)
                    except: pass
            
            popup.bind("<Button-1>", start_drag)
            popup.bind("<B1-Motion>", do_drag)
            
            content_frame = ttk.Frame(popup, padding=10)
            content_frame.pack(fill='both', expand=True)
            
            lb = tk.Listbox(content_frame, font=('Malgun Gothic', 10), 
                            bg='white', fg='#333333', relief='flat', 
                            borderwidth=0, highlightthickness=0)
            for info in full_info:
                lb.insert(tk.END, info)
            
            lb.pack(side='left', fill='both', expand=True)
            
            sb = ttk.Scrollbar(content_frame, orient="vertical", command=lb.yview)
            lb.configure(yscrollcommand=sb.set)
            sb.pack(side='right', fill='y')
            
            popup.focus_set()
            
            # --- Window Move Synchronization ---
            def reposition_popup(event=None):
                if not popup.winfo_exists(): return
                # Re-calculate position based on CURRENT root position + saved relative offset
                try:
                    curr_bbox = tree.bbox(item_id, column_id)
                    if not curr_bbox: return
                    cx, cy, cw, ch = curr_bbox
                    croot_x = tree.winfo_rootx() + cx
                    croot_y = tree.winfo_rooty() + cy
                    
                    # Apply relative offset to keep the sync accurate
                    new_x = croot_x + popup._rel_offset_x
                    new_y = croot_y + popup._rel_offset_y
                    
                    popup.geometry(f"+{int(new_x)}+{int(new_y)}")
                except: pass
                
            # Bind to root move/resize
            bind_id = self.root.bind("<Configure>", reposition_popup, add="+")
            
            def close_popup(e=None):
                if popup.winfo_exists():
                    try: self.root.unbind("<Configure>", bind_id)
                    except: pass
                    popup.destroy()
            
            popup.bind("<FocusOut>", close_popup)
            popup.bind("<Escape>", close_popup)
            popup.bind("<Double-1>", close_popup)
            
            # [FIX] Removed global self.root.bind("<Button-1>") which leaked and interfered with DateEntry.
            # FocusOut is sufficient for closing the popup when clicking elsewhere.
        except Exception as e:
            print(f"Final popup error: {e}")
    def clean_df_export(self, df):
        """Sanitize data for Excel: remove illegal characters, handle empty values, and protect formulas"""
        # 1. Pre-clean common empty representations
        df = df.replace(['nan', 'NaN', 'None'], "")
        
        # 2. Identify columns that have at least one meaningful value
        def is_really_empty(col):
            # A column is empty if all stripped values are "", NaN, or 0 sequences
            return df[col].astype(str).str.strip().replace(['nan', 'None', '', '0', '0.0', '0.00'], pd.NA).dropna().empty

        non_empty_cols = [col for col in df.columns if not is_really_empty(col)]
        
        # Always keep essential columns even if empty
        essential = ['날짜', '현장', '품목명', 'Date', 'Site']
        # Also keep standard RTK and NDT result columns
        rtk_cats = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타", "RTK총계"]
        ndt_materials = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
        essential.extend(rtk_cats)
        essential.extend([f'RTK_{c}' for c in rtk_cats])
        essential.extend(ndt_materials)
        essential.extend([f'NDT_{"".join(m.split())}' for m in ndt_materials])
        
        for col in essential:
            if col in df.columns and col not in non_empty_cols:
                non_empty_cols.append(col)
                
        # 3. Sanitize content and protect formulas
        def sanitize_and_escape(val):
            if pd.isna(val) or val == "" or val == 0 or val == 0.0:
                return ""
            
            # String conversion and cleaning
            s_val = str(val).strip()
            if not s_val or s_val.lower() in ['nan', 'none']:
                return ""

            # Remove illegal XML control characters (ASCII < 32 except tab, LF, CR)
            # These characters cause Excel to report file corruption.
            s_val = "".join(ch for ch in s_val if ord(ch) >= 32 or ch in "\t\n\r")
            
            # Escape potential formula characters (=, +, -, @)
            # IMPORTANT: Skip escaping if the value is a valid number (e.g., -5000)
            if s_val.startswith(('=', '+', '-', '@')):
                try:
                    # If it's a number, don't prefix with '
                    float(s_val)
                    return val
                except ValueError:
                    # It's a text string starting with a formula char
                    return f"'{s_val}"
            
            return s_val

        # Apply sanitization to all columns
        for col in df.columns:
            df[col] = df[col].apply(sanitize_and_escape)
        
        # Return only non-empty columns in original order
        final_cols = [c for c in df.columns if c in non_empty_cols]
        return df[final_cols]

    def save_df_to_excel_autofit(self, df, save_path, sheet_name='Sheet1'):
        """Save a DataFrame to Excel with automatic column width adjustment (AutoFit)"""
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            
            for idx, col in enumerate(df.columns):
                # Calculate max length of values in column + header
                # We handle Korean characters by assuming they take ~2 units of width
                def get_display_width(s):
                    width = 0
                    for char in str(s):
                        if ord(char) > 127: # Non-ASCII (Korean, etc.)
                            width += 2
                        else:
                            width += 1
                    return width

                series = df[col].astype(str)
                # Filter out empty strings/NAs for max calc
                lengths = series.apply(get_display_width)
                max_val_len = lengths.max() if not lengths.empty else 0
                header_len = get_display_width(col)
                
                # Final width with padding - Cap at a reasonable width to force wrapping
                final_width = min(max(max_val_len, header_len) + 2, 15)
                
                # Map index to column letter
                # column_letter property is available in openpyxl cells
                col_letter = worksheet.cell(row=1, column=idx+1).column_letter
                worksheet.column_dimensions[col_letter].width = final_width

            # --- Page Setup for Printing ---
            # Set to Portrait orientation (User request)
            worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
            
            # Enable Fit to Page width (Scale everything to fit horizontally)
            worksheet.sheet_properties.pageSetUpPr.fitToPage = True
            worksheet.page_setup.fitToWidth = 1
            worksheet.page_setup.fitToHeight = 0 # Automatic heights (multiple pages if long)
            
            # Enable Wrap Text for all cells to use vertical space instead of horizontal width
            from openpyxl.styles import Alignment
            wrap_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = wrap_alignment

            # Optional: Set small margins to maximize space (units are inches)
            worksheet.page_margins.left = 0.25
            worksheet.page_margins.right = 0.25
            worksheet.page_margins.top = 0.5
            worksheet.page_margins.bottom = 0.5


    
    def setup_import_tab(self):
        import_frame = ttk.LabelFrame(self.tab_import, text="데이터 관리")
        import_frame.pack(pady=20, padx=20, fill='both', expand=True)
        
        # Import Section
        ttk.Label(import_frame, text="엑셀 파일에서 자재 데이터 가져오기", font=('Arial', 11, 'bold')).pack(pady=10)
        ttk.Label(import_frame, text="형식: MaterialID, 회사코드, 관리품번, 품목명, 창고, 모델명, 규격, 품목군코드, 제조사, 제조국, 가격, 관리단위, 수량", 
                 wraplength=600).pack(pady=5)
        
        btn_import = ttk.Button(import_frame, text="엑셀 파일 가져오기", command=self.import_from_excel)
        btn_import.pack(pady=10)
        
        ttk.Separator(import_frame, orient='horizontal').pack(fill='x', pady=20)
        
        # Export Section
        ttk.Label(import_frame, text="현재 데이터 엑셀로 내보내기", font=('Arial', 11, 'bold')).pack(pady=10)
        
        btn_export_materials = ttk.Button(import_frame, text="자재 목록 내보내기", command=self.export_materials)
        btn_export_materials.pack(pady=5)
        
        btn_export_trans = ttk.Button(import_frame, text="거래 내역 내보내기", command=self.export_transactions)
        btn_export_trans.pack(pady=5)
        
        btn_export_all = ttk.Button(import_frame, text="전체 데이터 내보내기", command=self.export_all)
        btn_export_all.pack(pady=5)
    
    def import_from_excel(self):
        file_path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if not file_path:
            return
        
        try:
            imported_df = pd.read_excel(file_path)
            
            # Validate columns - accept both Korean and English column names
            required_cols = ['품목명'] if '품목명' in imported_df.columns else (['품명'] if '품명' in imported_df.columns else ['Item Name'])
            if not any(col in imported_df.columns for col in ['품목명', '품명', 'Item Name']):
                messagebox.showerror("오류", "필수 컬럼 '품목명', '품명' 또는 'Item Name'이 없습니다.")
                return
            
            # Process each row
            count_new = 0
            count_updated = 0
            
            # Ask user if they want to update existing items or just append
            update_existing = messagebox.askyesno("가져오기 방식", 
                "기존에 등록된 품목(품목명, SN, 규격 일치)이 있을 경우 정보를 업데이트하시겠습니까?\n\n"
                "'예': 기존 정보 수정\n'아니오': 무시하고 새로 추가")
            
            for _, row in imported_df.iterrows():
                # Extract values with mapping
                mat_name = str(row.get('품목명', row.get('품명', row.get('Item Name', '')))).strip()
                if not mat_name or mat_name == 'nan': continue
                
                sn = str(row.get('SN', row.get('SN번호', row.get('Serial Number', '')))).strip()
                if sn == 'nan': sn = ''
                
                spec = str(row.get('규격', row.get('Specification', ''))).strip()
                if spec == 'nan': spec = ''
                
                # Check for existing
                existing_idx = -1
                if not self.materials_df.empty:
                    # Match by Name, SN and Specification
                    mask = (self.materials_df['품목명'].astype(str) == mat_name) & \
                           (self.materials_df['SN'].astype(str).replace('nan', '') == sn) & \
                           (self.materials_df['규격'].astype(str).replace('nan', '') == spec)
                    
                    matches = self.materials_df.index[mask].tolist()
                    if matches:
                        existing_idx = matches[0]
                
                # Prepare data row
                extracted_model, extracted_sn = self.extract_sn_from_model(
                    row.get('모델명', row.get('Model', '')), 
                    sn
                )
                
                data_row = {
                    '회사코드': row.get('회사코드', row.get('Company Code', '')),
                    '관리품번': row.get('관리품번', row.get('Equipment Code', '')),
                    '품목명': mat_name,
                    'SN': extracted_sn,
                    '창고': row.get('창고', row.get('Warehouse', '')),
                    '모델명': extracted_model,
                    '규격': spec,
                    '품목군코드': row.get('품목군코드', row.get('Classification', '')),
                    '공급업체': row.get('공급업체', row.get('공급업자', row.get('Supplier', ''))),
                    '제조사': row.get('제조사', row.get('Manufacturer', '')),
                    '제조국': row.get('제조국', row.get('Country', row.get('Origin', ''))),
                    '가격': row.get('가격', row.get('Price', 0)),
                    '관리단위': row.get('관리단위', row.get('Unit', 'EA')),
                    '수량': row.get('수량', row.get('Initial Stock', row.get('Current Stock', 0))),
                    '재고하한': row.get('재고하한', row.get('재주문 수준', row.get('Reorder Point', 0)))
                }
                
                # Handle NaN values in numeric fields
                for field in ['가격', '수량', '재고하한']:
                    if pd.isna(data_row[field]): data_row[field] = 0
                
                if existing_idx != -1:
                    if update_existing:
                        # Update existing
                        for col, val in data_row.items():
                            self.materials_df.at[existing_idx, col] = val
                        count_updated += 1
                    else:
                        # Skip or append (here we append if skipping update to maintain old behavior but with new ID)
                        new_mat_id = self.materials_df['MaterialID'].max() + 1 if not self.materials_df.empty else 1
                        data_row['MaterialID'] = new_mat_id
                        self.materials_df = pd.concat([self.materials_df, pd.DataFrame([data_row])], ignore_index=True)
                        count_new += 1
                else:
                    # Add as new
                    new_mat_id = self.materials_df['MaterialID'].max() + 1 if not self.materials_df.empty else 1
                    data_row['MaterialID'] = new_mat_id
                    self.materials_df = pd.concat([self.materials_df, pd.DataFrame([data_row])], ignore_index=True)
                    count_new += 1
            
            self.save_data()
            self.update_material_combo()
            self.update_stock_view()
            self.update_registration_combos()
            
            msg = f"자재 가져오기가 완료되었습니다.\n\n"
            if count_new > 0: msg += f"• 신규 등록: {count_new}건\n"
            if count_updated > 0: msg += f"• 기존 수정: {count_updated}건"
            messagebox.showinfo("완료", msg)
            
        except Exception as e:
            messagebox.showerror("오류", f"파일을 가져오는데 실패했습니다: {e}")
    
    def export_materials(self):
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="Materials_Export.xlsx",
            title="자재 목록 저장",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if save_path:
            try:
                self.materials_df.to_excel(save_path, index=False)
                messagebox.showinfo("완료", "자재 목록이 저장되었습니다.")
            except Exception as e:
                messagebox.showerror("오류", f"저장 실패: {e}")
    
    def export_transactions(self):
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="Transactions_Export.xlsx",
            title="거래 내역 저장",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if save_path:
            try:
                self.transactions_df.to_excel(save_path, index=False)
                messagebox.showinfo("완료", "거래 내역이 저장되었습니다.")
            except Exception as e:
                messagebox.showerror("오류", f"저장 실패: {e}")
    
    def export_transaction_history(self):
        """Export transaction history displayed in the inout_tree to Excel"""
        # Build data from current treeview
        history_data = []
        for item in self.inout_tree.get_children():
            values = self.inout_tree.item(item, 'values')
            history_data.append({
                '날짜': values[0],
                '현장': values[1],
                '품목명': values[2],
                '모델명': values[3],
                '구분': values[4],
                '수량': values[5],
                '재고': values[6],
                '담당자': values[7],
                '차량번호': values[8],
                '주행거리': values[9],
                '차량점검': values[10],
                '차량비고': values[11],
                '비고': values[12]
            })
        
        if not history_data:
            messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
            return

    def _migrate_worktimes(self, input_list):
        """Helper to standardize work times with 익일 marker and sort logically"""
        default_worktimes = [
            "09:00~18:00", "09:00~19:00", "09:00~20:00", "09:00~21:00", 
            "09:00~22:00", "09:00~23:00", "09:00~24:00", "09:00~익일01:00",
            "09:00~익일02:00", "09:00~익일03:00", "18:00~익일02:00", "18:00~익일03:00"
        ]
        
        processed_times = set()
        for t in list(input_list) + default_worktimes:
            if not t or '~' not in t: continue
            
            clean_t = str(t).replace('익일', '').strip()
            try:
                start, end = clean_t.split('~')
                s_h = int(start.split(':')[0])
                e_h = int(end.split(':')[0])
                
                if e_h < s_h and '익일' not in str(t):
                    processed_times.add(f"{start}~익일{end}")
                else:
                    processed_times.add(str(t).strip())
            except:
                processed_times.add(str(t).strip())

        def time_sort_key(t):
            has_ikil = 1 if '익일' in t else 0
            clean = t.replace('익일', '')
            return (has_ikil, clean)

        return sorted(list(processed_times), key=time_sort_key)
        
        # Prepare filename
        today = datetime.datetime.now().strftime('%Y%m%d')
        filename = f"입출고내역_{today}.xlsx"
        
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=filename,
            title="입출고 내역 저장",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if save_path:
            try:
                history_df = pd.DataFrame(history_data)
                history_df = self.clean_df_export(history_df)
                self.save_df_to_excel_autofit(history_df, save_path, "입출고내역")
                messagebox.showinfo("완료", "입출고 내역이 저장되었습니다.")
            except Exception as e:
                messagebox.showerror("오류", f"저장 실패: {e}")
    
    def export_all(self):
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="Complete_Export.xlsx",
            title="전체 데이터 저장",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if save_path:
            try:
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    self.materials_df.to_excel(writer, sheet_name='Materials', index=False)
                    self.transactions_df.to_excel(writer, sheet_name='Transactions', index=False)
                messagebox.showinfo("완료", "전체 데이터가 저장되었습니다.")
            except Exception as e:
                messagebox.showerror("오류", f"저장 실패: {e}")
    

    def view_monthly_usage(self):
        """Display monthly usage by item name in the treeview"""
        # Clear current view
        for item in self.usage_tree.get_children():
            self.usage_tree.delete(item)
        
        year = int(self.cb_year.get())
        month = int(self.cb_month.get())
        
        # Filter transactions for the selected month
        month_mask = (self.transactions_df['Date'].dt.year == year) & \
                     (self.transactions_df['Date'].dt.month == month) & \
                     (self.transactions_df['Type'] == 'OUT')
        monthly_trans = self.transactions_df[month_mask]
        
        # Filter transactions for cumulative (from start of year to selected month)
        cumulative_mask = (self.transactions_df['Date'].dt.year == year) & \
                          (self.transactions_df['Date'].dt.month <= month) & \
                          (self.transactions_df['Type'] == 'OUT')
        cumulative_trans = self.transactions_df[cumulative_mask]
        
        # Build usage data for each material
        usage_data = []
        for _, mat in self.materials_df.iterrows():
            mat_id = mat['MaterialID']
            
            # Calculate monthly usage
            month_usage = monthly_trans[monthly_trans['MaterialID'] == mat_id]['Quantity'].sum()
            
            # Calculate cumulative usage (year-to-date)
            cumulative_usage = cumulative_trans[cumulative_trans['MaterialID'] == mat_id]['Quantity'].sum()
            
            # Only show items with usage
            if month_usage > 0 or cumulative_usage > 0:
                usage_data.append({
                    '품목명': mat.get('품목명', ''),  # Using 재고현황's 품목명 field
                    '관리품번': mat.get('관리품번', ''),
                    '규격': mat.get('규격', ''),
                    '단위': mat.get('관리단위', 'EA'),
                    '월사용량': month_usage,
                    '누계사용량': cumulative_usage
                })
        
        # Sort by item name
        usage_data.sort(key=lambda x: x['품목명'])
        
        # Display in treeview
        for data in usage_data:
            self.usage_tree.insert('', tk.END, values=(
                data['품목명'],
                data['관리품번'],
                data['규격'],
                data['단위'],
                f"{data['월사용량']:.1f}",
                f"{data['누계사용량']:.1f}"
            ))
        
        # Show message if no data
        if not usage_data:
            messagebox.showinfo("알림", f"{year}년 {month}월에 사용 내역이 없습니다.")

    def generate_yearly_report(self):
        year = int(self.cb_year.get())
        # Filter transactions for the year
        mask = (self.transactions_df['Date'].dt.year == year) & (self.transactions_df['Type'] == 'OUT')
        yearly_trans = self.transactions_df[mask]
        
        report = []
        for _, mat in self.materials_df.iterrows():
            mat_id = mat['MaterialID']
            
            # Calculate monthly usage for each month
            row_data = {
                '설비코드': mat.get('관리품번', ''),
                '자재명': mat.get('품목명', ''),  # Using 재고현황's 품목명 field
                '분류': mat.get('품목군코드', ''),
                '규격': mat.get('규격', ''),
                '단위': mat.get('관리단위', ''),
                '제조사': mat.get('제조사', '')
            }
            
            # Add monthly columns (1월 ~ 12월)
            monthly_values = []
            for month in range(1, 13):
                month_mask = (yearly_trans['MaterialID'] == mat_id) & \
                            (yearly_trans['Date'].dt.month == month)
                month_usage = yearly_trans[month_mask]['Quantity'].sum()
                row_data[f'{month}월'] = month_usage
                monthly_values.append(month_usage)
            
            # Calculate totals
            total = sum(monthly_values)
            row_data['합계'] = total
            
            # Calculate cumulative total
            cumulative = 0
            for i, val in enumerate(monthly_values, 1):
                cumulative += val
                if i == 12:  # Only show final cumulative at the end
                    row_data['누계'] = cumulative
            
            report.append(row_data)
            
        report_df = pd.DataFrame(report)
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                 initialfile=f"Yearly_Usage_{year}.xlsx",
                                                 title="보고서 저장")
        if save_path:
            report_df.to_excel(save_path, index=False)
            messagebox.showinfo("완료", f"{year}년 보고서가 저장되었습니다.")

    def generate_monthly_report(self):
        year = int(self.cb_year.get())
        month = int(self.cb_month.get())
        
        mask = (self.transactions_df['Date'].dt.year == year) & \
               (self.transactions_df['Date'].dt.month == month) & \
               (self.transactions_df['Type'] == 'OUT')
        monthly_trans = self.transactions_df[mask]
        
        report = []
        for _, mat in self.materials_df.iterrows():
            mat_id = mat['MaterialID']
            total_usage = monthly_trans[monthly_trans['MaterialID'] == mat_id]['Quantity'].sum()
            report.append({
                '설비코드': mat.get('관리품번', ''),
                '자재명': mat.get('품목명', ''),  # Using 재고현황's 품목명 field
                '분류': mat.get('품목군코드', ''),
                '규격': mat.get('규격', ''),
                '단위': mat.get('관리단위', ''),
                '제조사': mat.get('제조사', ''),
                '월간 총 사용량': total_usage
            })
            
        report_df = pd.DataFrame(report)
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                 initialfile=f"Monthly_Usage_{year}_{month}.xlsx",
                                                 title="보고서 저장")
        if save_path:
            report_df.to_excel(save_path, index=False)
            messagebox.showinfo("완료", f"{year}년 {month}월 보고서가 저장되었습니다.")

    def setup_monthly_usage_tab(self):
        """Setup the monthly usage aggregation tab (auto-aggregated from daily usage)"""
        # Display frame for aggregated monthly data
        display_frame = ttk.LabelFrame(self.tab_monthly_usage, text="월별 사용량 집계 (현장별 데이터 자동 집계)")
        display_frame.pack(expand=True, fill='both', padx=10, pady=10)
        
        # Filter controls
        filter_frame = ttk.Frame(display_frame)
        filter_frame.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(filter_frame, text="연도:").pack(side='left', padx=5)
        current_year = datetime.datetime.now().year
        year_values = ['전체'] + [str(y) for y in range(max(2024, current_year-1), current_year + 3)]
        self.cb_filter_year = ttk.Combobox(filter_frame, values=year_values, width=10)
        self.cb_filter_year.pack(side='left', padx=5)
        self.cb_filter_year.set(str(current_year))
        
        ttk.Label(filter_frame, text="월:").pack(side='left', padx=5)
        current_month = datetime.datetime.now().month
        self.cb_filter_month = ttk.Combobox(filter_frame, values=['전체'] + [str(m) for m in range(1, 13)], width=10)
        self.cb_filter_month.pack(side='left', padx=5)
        self.cb_filter_month.set(str(current_month))
        
        ttk.Label(filter_frame, text="현장:").pack(side='left', padx=5)
        self.cb_filter_site_monthly = ttk.Combobox(filter_frame, width=15)
        self.cb_filter_site_monthly.pack(side='left', padx=5)
        self.cb_filter_site_monthly.set('전체')
        
        ttk.Label(filter_frame, text="품목명:").pack(side='left', padx=5)
        self.cb_filter_material_monthly = ttk.Combobox(filter_frame, width=25)
        self.cb_filter_material_monthly.pack(side='left', padx=5)
        self.cb_filter_material_monthly.set('전체')
        
        btn_filter = ttk.Button(filter_frame, text="조회", command=self.update_monthly_usage_view)
        btn_filter.pack(side='left', padx=10)
        
        btn_export = ttk.Button(filter_frame, text="엑셀 내보내기", command=self.export_monthly_usage_history)
        btn_export.pack(side='left', padx=5)
        
        # Use a PanedWindow to allow resizing between main tree and summaries
        self.monthly_paned = ttk.PanedWindow(display_frame, orient="vertical")
        self.monthly_paned.pack(expand=True, fill='both', padx=5, pady=5)
        
        # 1. Top pane: Main Monthly Usage Tree
        tree_frame = ttk.Frame(self.monthly_paned)
        self.monthly_paned.add(tree_frame, weight=3) # Give main tree more weight
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        # Treeview with columns including worker, work time, and OT fields
        # Added '(Full작업자)' for full list backup during Excel export
        columns = ('연도', '월', '현장', '작업자', '작업시간', 'OT시간', 'OT금액', 'OT1', 'OT2', 'OT3', 'OT4', 'OT5', 'OT6', 'OT7', 'OT8', 'OT9', 'OT10', 
                   '수량', '단가', '출장비', '일식', '검사비', '품목명', '필름매수', '센터미스', '농도', '마킹미스', '필름마크', 
                   '취급부주의', '고객불만', '기타', 'RTK총계', '형광자분', '흑색자분', '백색페인트', '침투제', '세척제', '현상제', '형광침투제', '비고', '(Full작업자)')
        self.monthly_usage_tree = ttk.Treeview(tree_frame, columns=columns, show='headings',
                                               yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Hide the (Full작업자) column from display
        self.monthly_usage_tree['displaycolumns'] = [c for c in columns if c != '(Full작업자)']
        
        vsb.config(command=self.monthly_usage_tree.yview)
        hsb.config(command=self.monthly_usage_tree.xview)
        
        # Column configuration with added OT columns
        col_widths = {
            '연도': 90, '월': 70, '현장': 140, '작업자': 100, '작업시간': 100,
            'OT시간': 100, 'OT금액': 110,
            'OT1': 100, 'OT2': 100, 'OT3': 100, 'OT4': 100, 'OT5': 100,
            'OT6': 100, 'OT7': 100, 'OT8': 100, 'OT9': 100, 'OT10': 100,
            '수량': 110, '단가': 110, '출장비': 110, '일식': 110, '검사비': 110,
            '품목명': 220, '필름매수': 110, '센터미스': 80, '농도': 80, '마킹미스': 80,
            '필름마크': 80, '취급부주의': 80, '고객불만': 80, '기타': 80, 'RTK총계': 80,
            '형광자분': 90, '흑색자분': 90, '백색페인트': 90, '침투제': 90, '세척제': 90,
            '현상제': 90, '형광침투제': 90, '비고': 220
        }
        
        for col in columns:
            self.monthly_usage_tree.heading(col, text=col, command=lambda c=col: self.treeview_sort_column(self.monthly_usage_tree, c, False))
            width = col_widths.get(col, 100)
            self.monthly_usage_tree.column(col, width=width, minwidth=20, stretch=False, anchor='center')
        
        # [NEW] Enable column reordering via drag & drop
        self.enable_tree_column_drag(self.monthly_usage_tree)
        
        
        # Grid layout
        self.monthly_usage_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # [NEW] Auto-save column widths when user resizes columns
        self.monthly_usage_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())
        
        # [NEW] Bind selection to update site/worker summaries
        self.monthly_usage_tree.bind('<<TreeviewSelect>>', self.on_monthly_usage_select)

        # 2. Middle pane: Site Summary
        site_frame = ttk.LabelFrame(self.monthly_paned, text="현장별 누계")
        self.monthly_paned.add(site_frame, weight=1)
        
        self.monthly_usage_tree.bind("<Button-1>", lambda e: self.show_worker_popup(e, self.monthly_usage_tree), add="+")
        
        site_cols = ('현장', '품목명', '필름매수', '검사비', '출장비', '형광자분', '흑색자분', '백색페인트', 
                     '침투제', '세척제', '현상제', '형광침투제', '센터미스', '농도', '마킹미스', '필름마크', '취급부주의', '고객불만', '기타', 'RTK총계')
        self.site_summary_tree = ttk.Treeview(site_frame, columns=site_cols, show='headings')
        for col in site_cols:
            self.site_summary_tree.heading(col, text=col)
            # Adjust widths based on content
            if col in ['현장', '품목명']: width = 120
            elif col in ['검사비', '출장비']: width = 100
            else: width = 80
            self.site_summary_tree.column(col, width=width, anchor='center', stretch=False)
        
        # [NEW] Auto-save site summary column widths
        self.site_summary_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())
        
        site_vsb = ttk.Scrollbar(site_frame, orient="vertical", command=self.site_summary_tree.yview)
        self.site_summary_tree.configure(yscrollcommand=site_vsb.set)
        
        self.site_summary_tree.pack(side='left', expand=True, fill='both')
        site_vsb.pack(side='right', fill='y')
        
        # 3. Bottom pane: Worker Summary
        worker_frame = ttk.LabelFrame(self.monthly_paned, text="작업자별 누계")
        self.monthly_paned.add(worker_frame, weight=1)
        
        worker_cols = ('작업자', '총공수', '연장(시간)', '야간(시간)', '휴일(시간)', '총OT(시간)', '연장(금액)', '야간(금액)', '휴일(금액)', '총OT(금액)')
        self.worker_summary_tree = ttk.Treeview(worker_frame, columns=worker_cols, show='headings')
        
        # Set column widths for worker summary
        worker_widths = {
            '작업자': 100, '총공수': 70, '연장(시간)': 70, '야간(시간)': 70, '휴일(시간)': 70, 
            '총OT(시간)': 80, '연장(금액)': 90, '야간(금액)': 90, '휴일(금액)': 90, '총OT(금액)': 100
        }
        for col in worker_cols:
            self.worker_summary_tree.heading(col, text=col)
            width = worker_widths.get(col, 80)
            # Disable stretching
            self.worker_summary_tree.column(col, width=width, anchor='center', stretch=False)
        
        # [NEW] Auto-save worker summary column widths
        self.worker_summary_tree.bind('<ButtonRelease-1>', lambda e: self.save_tab_config())
            
        worker_vsb = ttk.Scrollbar(worker_frame, orient="vertical", command=self.worker_summary_tree.yview)
        self.worker_summary_tree.configure(yscrollcommand=worker_vsb.set)
        
        self.worker_summary_tree.pack(side='left', expand=True, fill='both')
        worker_vsb.pack(side='right', fill='y')

        # [NEW] 공사탭 특별근무 자동 입력 버튼
        btn_apply_frame = ttk.Frame(worker_frame)
        btn_apply_frame.pack(fill='x', pady=3)
        ttk.Button(btn_apply_frame,
                   text="📋 공사탭 특별근무에 적용",
                   command=self.apply_worker_shift_hours_to_budget).pack(side='right', padx=5)

        # Initial view update
        self.update_monthly_usage_view()
    
    def update_monthly_usage_view(self):
        """Update the monthly usage treeview with aggregated data from daily usage"""
        # Clear current views
        for item in self.monthly_usage_tree.get_children():
            self.monthly_usage_tree.delete(item)
        for item in self.site_summary_tree.get_children():
            self.site_summary_tree.delete(item)
        for item in self.worker_summary_tree.get_children():
            self.worker_summary_tree.delete(item)
        
        # Get filter values
        filter_year = self.cb_filter_year.get()
        filter_month = self.cb_filter_month.get()
        filter_site = self.cb_filter_site_monthly.get() if hasattr(self, 'cb_filter_site_monthly') else '전체'
        filter_material = self.cb_filter_material_monthly.get() if hasattr(self, 'cb_filter_material_monthly') else '전체'
        
        # Return if daily usage data is empty
        if self.daily_usage_df.empty:
            return
        
        # Create a copy of daily usage data and extract year/month from Date column
        df = self.daily_usage_df.copy()
        
        # Normalize column names - remove ALL types of whitespace using regex
        import re
        df.columns = [re.sub(r'\s+', '', str(c)) for c in df.columns]
        
        df['Year'] = pd.to_datetime(df['Date'], errors='coerce').dt.year
        df['Month'] = pd.to_datetime(df['Date'], errors='coerce').dt.month
        # 날짜 파싱 실패 행 제거
        df = df.dropna(subset=['Year', 'Month'])
        df['Year'] = df['Year'].astype(int)
        df['Month'] = df['Month'].astype(int)
        
        # Apply filters
        if filter_year != '전체':
            df = df[df['Year'] == int(filter_year)]
        
        if filter_month != '전체':
            df = df[df['Month'] == int(filter_month)]
        
        if filter_site != '전체':
            df = df[df['Site'] == filter_site]
        
        # [NEW] Return early if filtered df is empty to avoid ValueError during assignment
        if df.empty:
            return
        # Populate site filter options from data
        if hasattr(self, 'cb_filter_site_monthly'):
            unique_sites = ['전체'] + sorted(self.daily_usage_df['Site'].dropna().unique().tolist())
            self.cb_filter_site_monthly['values'] = unique_sites
            if not self.cb_filter_site_monthly.get():
                self.cb_filter_site_monthly.set('전체')
        
        # Populate material filter options from data
        if hasattr(self, 'cb_filter_material_monthly'):
            # Get unique material names from materials_df based on MaterialIDs in daily_usage_df
            # Note: MaterialID column name itself might have spaces in self.daily_usage_df
            m_id_col = 'MaterialID' if 'MaterialID' in self.daily_usage_df.columns else 'MaterialID'
            unique_mat_ids = self.daily_usage_df[m_id_col].dropna().unique()
            material_names = []
            for mat_id in unique_mat_ids:
                mat_row = self.materials_df[self.materials_df['MaterialID'] == mat_id]
                if not mat_row.empty:
                    material_names.append(mat_row.iloc[0]['품목명'])
            unique_materials = ['전체'] + sorted(set(material_names))
            self.cb_filter_material_monthly['values'] = unique_materials
            if not self.cb_filter_material_monthly.get():
                self.cb_filter_material_monthly.set('전체')
        
        # Prepare aggregation dictionary for all numeric fields
        agg_dict = {'Usage': 'sum'}
        
        # Helper for joining workers - clean internal spaces to avoid "11시간" vs "11 시간" mismatch
        def join_unique_non_empty(series):
            # Strip outer spaces and compress internal spaces for consistency
            vals = [" ".join(str(v).split()) for v in series if pd.notna(v) and str(v).strip()]
            return " | ".join(sorted(set(vals)))
        
        # Also define a sum helper that handles potential type issues
        def safe_sum(series):
            return pd.to_numeric(series, errors='coerce').sum()

        # Add worker info - combine unique values from User, User2, ..., User10
        worker_cols = ['User', 'User2', 'User3', 'User4', 'User5', 'User6', 'User7', 'User8', 'User9', 'User10']
        for col in worker_cols:
            if col in df.columns:
                agg_dict[col] = join_unique_non_empty
        
        # Add work time - join unique values
        worktime_cols = ['WorkTime', 'WorkTime2', 'WorkTime3', 'WorkTime4', 'WorkTime5', 'WorkTime6', 'WorkTime7', 'WorkTime8', 'WorkTime9', 'WorkTime10']
        for col in worktime_cols:
            if col in df.columns:
                agg_dict[col] = join_unique_non_empty
        
        # Add OT - join unique values
        ot_cols = ['OT', 'OT2', 'OT3', 'OT4', 'OT5', 'OT6', 'OT7', 'OT8', 'OT9', 'OT10']
        for col in ot_cols:
            if col in df.columns:
                agg_dict[col] = join_unique_non_empty
        
        # Add Film Count if exists
        if 'FilmCount' in df.columns:
            agg_dict['FilmCount'] = safe_sum
        
        # Add RTK categories
        rtk_categories = ['RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크', 'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타']
        for cat in rtk_categories:
            if cat in df.columns:
                agg_dict[cat] = safe_sum
        
        # Add NDT materials and cost fields
        other_agg_cols = ['NDT_형광자분', 'NDT_자분', 'NDT_흑색자분', 'NDT_페인트', 'NDT_백색페인트', 'NDT_침투제', 'NDT_세척제', 'NDT_현상제', 'NDT_형광', 'NDT_형광침투제',
                          '검사량', '단가', '출장비', '일식', '검사비']
        for col in other_agg_cols:
            if col in df.columns:
                agg_dict[col] = safe_sum

        # Add calculated OT and WorkerCount columns for per-row aggregation
        def calc_row_values(row):
            h_sum = 0
            a_sum = 0
            w_count = 0
            for i in range(1, 11):
                # OT calculation
                ot_c = 'OT' if i == 1 else f'OT{i}'
                if ot_c in row and pd.notna(row[ot_c]):
                    v = str(row[ot_c]).strip()
                    if v:
                        h_sum += self._parse_ot_hours(v)
                        a_sum += self.calculate_ot_amount(v)
                
                # Worker count (Man-power)
                u_c = 'User' if i == 1 else f'User{i}'
                if u_c in row and pd.notna(row[u_c]):
                    u_val = str(row[u_c]).strip()
                    if u_val and u_val != 'nan' and u_val != '0.0':
                        w_count += 1
                        
            return pd.Series([h_sum, a_sum, w_count])

        df[['OT시간', 'OT금액', 'WorkerCount']] = df.apply(calc_row_values, axis=1)
        agg_dict['OT시간'] = 'sum'
        agg_dict['OT금액'] = 'sum'
        agg_dict['WorkerCount'] = 'sum'

        # Group by Year, Month, Site, MaterialID and aggregate
        grouped = df.groupby(['Year', 'Month', 'Site', 'MaterialID']).agg(agg_dict).reset_index()
        
        self.monthly_usage_tree.bind("<Button-1>", lambda e: self.show_worker_popup(e, self.monthly_usage_tree), add="+")
        
        # Store df for selection handling
        self.current_monthly_df = df
        
        # Initialize totals for cumulative sum
        total_worker_count = 0.0
        total_ot_hours = 0.0
        total_ot_amount = 0.0
        total_film_count = 0.0
        total_test_amount = 0.0
        total_unit_price = 0.0
        total_travel_cost = 0.0
        total_meal_cost = 0.0
        total_test_fee = 0.0
        total_rtk_center = 0.0
        total_rtk_density = 0.0
        total_rtk_marking = 0.0
        total_rtk_film = 0.0
        total_rtk_handling = 0.0
        total_rtk_customer = 0.0
        total_rtk_other = 0.0
        total_ndt_fluorescent_mag = 0.0
        total_ndt_magnet = 0.0
        total_ndt_paint = 0.0
        total_ndt_penetrant = 0.0
        total_ndt_cleaner = 0.0
        total_ndt_developer = 0.0
        total_ndt_fluorescent_pen = 0.0
        total_indiv_ot_amounts = [0] * 10
        has_note = False
        
        # Display aggregated entries
        for _, entry in grouped.iterrows():
            def clean_str(val):
                return str(val).replace('nan', '').replace('None', '').strip()
                
            mat_id = entry['MaterialID']
            mat_row = self.materials_df[self.materials_df['MaterialID'] == mat_id]
            
            if not mat_row.empty:
                mat_name = mat_row.iloc[0]['품목명']
            else:
                mat_name = f"ID: {mat_id}"
            
            # Apply material filter
            if filter_material != '전체' and mat_name != filter_material:
                continue
            
            # Get aggregated values from clean columns
            work_count = entry.get('WorkerCount', 0.0)
            ot_hours = entry.get('OT시간', 0.0)
            ot_amount = entry.get('OT금액', 0.0)
            film_count = entry.get('FilmCount', 0.0)
            test_amount = entry.get('검사량', 0.0)
            unit_price = entry.get('단가', 0.0)
            travel_cost = entry.get('출장비', 0.0)
            meal_cost = entry.get('일식', 0.0)
            test_fee = entry.get('검사비', 0.0)
            
            # RTK values
            rtk_center = entry.get('RTK_센터미스', 0.0)
            rtk_density = entry.get('RTK_농도', 0.0)
            rtk_marking = entry.get('RTK_마킹미스', 0.0)
            rtk_film = entry.get('RTK_필름마크', 0.0)
            rtk_handling = entry.get('RTK_취급부주의', 0.0)
            rtk_customer = entry.get('RTK_고객불만', 0.0)
            rtk_other = entry.get('RTK_기타', 0.0)
            rtk_total = rtk_center + rtk_density + rtk_marking + rtk_film + rtk_handling + rtk_customer + rtk_other
            
            # NDT values
            ndt_fluorescent_mag = entry.get('NDT_형광자분', 0.0)
            ndt_magnet = entry.get('NDT_자분', 0.0) + entry.get('NDT_흑색자분', 0.0)
            ndt_paint = entry.get('NDT_페인트', 0.0) + entry.get('NDT_백색페인트', 0.0)
            ndt_penetrant = entry.get('NDT_침투제', 0.0)
            ndt_cleaner = entry.get('NDT_세척제', 0.0)
            ndt_developer = entry.get('NDT_현상제', 0.0)
            ndt_fluorescent_pen = entry.get('NDT_형광', 0.0) + entry.get('NDT_형광침투제', 0.0)
            
            # Accumulate totals
            total_worker_count += work_count
            total_ot_hours += ot_hours
            total_ot_amount += ot_amount
            total_film_count += film_count
            total_test_amount += test_amount
            total_unit_price += unit_price
            total_travel_cost += travel_cost
            total_meal_cost += meal_cost
            total_test_fee += test_fee
            
            total_rtk_center += rtk_center
            total_rtk_density += rtk_density
            total_rtk_marking += rtk_marking
            total_rtk_film += rtk_film
            total_rtk_handling += rtk_handling
            total_rtk_customer += rtk_customer
            total_rtk_other += rtk_other
            total_ndt_fluorescent_mag += ndt_fluorescent_mag
            total_ndt_magnet += ndt_magnet
            total_ndt_paint += ndt_paint
            total_ndt_penetrant += ndt_penetrant
            total_ndt_cleaner += ndt_cleaner
            total_ndt_developer += ndt_developer
            total_ndt_fluorescent_pen += ndt_fluorescent_pen
            
            # Extract worker names from User, User2, ..., User10
            all_workers = []
            worker_cols = ['User', 'User2', 'User3', 'User4', 'User5', 'User6', 'User7', 'User8', 'User9', 'User10']
            for col in worker_cols:
                val = str(entry.get(col, '')).strip()
                if val and val.lower() not in ['nan', '0.0', 'none']:
                    # Split in case it was already joined in aggregation
                    all_workers.extend([v.strip() for v in val.split(' | ') if v.strip()])
            worker_str = self.format_worker_summary(all_workers)
            
            # Concatenate work times
            all_worktimes = []
            worktime_cols = ['WorkTime', 'WorkTime2', 'WorkTime3', 'WorkTime4', 'WorkTime5', 'WorkTime6', 'WorkTime7', 'WorkTime8', 'WorkTime9', 'WorkTime10']
            for col in worktime_cols:
                val = str(entry.get(col, '')).strip()
                if val and val != 'nan' and val != '0.0':
                    all_worktimes.extend([v.strip() for v in val.split(' | ') if v.strip()])
            worktime_str = ", ".join(sorted(set(all_worktimes)))
            
            # Extract OT amounts only (suppress time strings)
            ot_values = []
            ot_cols = ['OT', 'OT2', 'OT3', 'OT4', 'OT5', 'OT6', 'OT7', 'OT8', 'OT9', 'OT10']
            for col in ot_cols:
                val = str(entry.get(col, '')).strip()
                if val and val != 'nan' and val != '0.0':
                    # Handle multiple OTs if joined by aggregation separator ' | '
                    sub_vals = [v.strip() for v in val.split(' | ') if v.strip()]
                    parsed_ots = []
                    for v_str in sub_vals:
                        if '(' in v_str and '원)' in v_str:
                            try:
                                amount_str = v_str.split('(')[1].split('원')[0].replace(',', '').strip()
                                amount = int(float(amount_str))
                                parsed_ots.append(f"{amount:,}")
                            except:
                                pass # Skip invalid format
                        else:
                            # Try to calculate amount from time-like string
                            try:
                                amt = self.calculate_ot_amount(v_str)
                                if amt > 0:
                                    parsed_ots.append(f"{amt:,}")
                                # If it looks like time (contains ':', '시', '시간', '~', '-') but amt is 0, hide it
                                elif any(x in v_str for x in [':', '시', '시간', '~', '-']):
                                    pass 
                                else:
                                    parsed_ots.append(v_str) # Keep non-time comments
                            except:
                                pass
                    ot_values.append(", ".join(parsed_ots))
                    if any(v.replace(',', '').isnumeric() for v in parsed_ots):
                        try:
                            # Sum only the first numeric value for status check
                            amt = int(parsed_ots[0].replace(',', ''))
                            total_indiv_ot_amounts[i-1] += amt
                        except: pass
                else:
                    ot_values.append('')
            
            if clean_str(entry.get('Note', '')): has_note = True
            
            self.monthly_usage_tree.insert('', tk.END, values=(
                int(entry['Year']),
                int(entry['Month']),
                entry.get('Site', ''),
                worker_str,  # 작업자
                worktime_str,  # 작업시간 (concatenated strings)
                f"{ot_hours:.1f}" if ot_hours > 0 else '',
                f"{ot_amount:,.0f}" if ot_amount > 0 else '',
                *ot_values,  # OT1-10 (parsed amounts)
                f"{test_amount:.1f}" if test_amount > 0 else '',
                f"{unit_price:,.0f}" if unit_price > 0 else '',
                f"{travel_cost:,.0f}" if travel_cost > 0 else '',
                f"{meal_cost:,.0f}" if meal_cost > 0 else '',
                f"{test_fee:,.0f}" if test_fee > 0 else '',
                mat_name,
                f"{film_count:.1f}" if film_count > 0 else '',
                f"{rtk_center:.1f}" if rtk_center > 0 else '',
                f"{rtk_density:.1f}" if rtk_density > 0 else '',
                f"{rtk_marking:.1f}" if rtk_marking > 0 else '',
                f"{rtk_film:.1f}" if rtk_film > 0 else '',
                f"{rtk_handling:.1f}" if rtk_handling > 0 else '',
                f"{rtk_customer:.1f}" if rtk_customer > 0 else '',
                f"{rtk_other:.1f}" if rtk_other > 0 else '',
                f"{rtk_total:.1f}" if rtk_total > 0 else '',
                f"{ndt_fluorescent_mag:.1f}" if ndt_fluorescent_mag > 0 else '',
                f"{ndt_magnet:.1f}" if ndt_magnet > 0 else '',
                f"{ndt_paint:.1f}" if ndt_paint > 0 else '',
                f"{ndt_penetrant:.1f}" if ndt_penetrant > 0 else '',
                f"{ndt_cleaner:.1f}" if ndt_cleaner > 0 else '',
                f"{ndt_developer:.1f}" if ndt_developer > 0 else '',
                f"{ndt_fluorescent_pen:.1f}" if ndt_fluorescent_pen > 0 else '',
                '',  # Empty note field
                ", ".join(sorted(set(all_workers))) # (Full작업자)
            ))
        
        # Add total row at the bottom if there's data
        if not grouped.empty:
            total_rtk_sum = total_rtk_center + total_rtk_density + total_rtk_marking + total_rtk_film + total_rtk_handling + total_rtk_customer + total_rtk_other
            
            # Configure tag for total row
            self.monthly_usage_tree.tag_configure('total', background='#E8F4F8', font=('Arial', 12, 'bold'))
            
            total_values = (
                '',
                '',
                '--- 전체 누계 ---',
                '',  # 작업자
                f"{total_worker_count:.1f}" if total_worker_count > 0 else '', # 작업시간 (총 공수)
                f"{total_ot_hours:.1f}" if total_ot_hours > 0 else '',
                f"{total_ot_amount:,.0f}" if total_ot_amount > 0 else '',
                '', '', '', '', '', '', '', '', '', '',  # OT1-10
                f"{total_test_amount:,.0f}" if total_test_amount > 0 else '',
                '', # 단가
                f"{total_travel_cost:,.0f}" if total_travel_cost > 0 else '',
                f"{total_meal_cost:,.0f}" if total_meal_cost > 0 else '',
                f"{total_test_fee:,.0f}" if total_test_fee > 0 else '',
                '', # 품목명
                f"{total_film_count:.1f}" if total_film_count > 0 else '',
                f"{total_rtk_center:.1f}" if total_rtk_center > 0 else '',
                f"{total_rtk_density:.1f}" if total_rtk_density > 0 else '',
                f"{total_rtk_marking:.1f}" if total_rtk_marking > 0 else '',
                f"{total_rtk_film:.1f}" if total_rtk_film > 0 else '',
                f"{total_rtk_handling:.1f}" if total_rtk_handling > 0 else '',
                f"{total_rtk_customer:.1f}" if total_rtk_customer > 0 else '',
                f"{total_rtk_other:.1f}" if total_rtk_other > 0 else '',
                f"{total_rtk_sum:.1f}" if total_rtk_sum > 0 else '',
                f"{total_ndt_fluorescent_mag:.1f}" if total_ndt_fluorescent_mag > 0 else '',
                f"{total_ndt_magnet:.1f}" if total_ndt_magnet > 0 else '',
                f"{total_ndt_paint:.1f}" if total_ndt_paint > 0 else '',
                f"{total_ndt_penetrant:.1f}" if total_ndt_penetrant > 0 else '',
                f"{total_ndt_cleaner:.1f}" if total_ndt_cleaner > 0 else '',
                f"{total_ndt_developer:.1f}" if total_ndt_developer > 0 else '',
                f"{total_ndt_fluorescent_pen:.1f}" if total_ndt_fluorescent_pen > 0 else '',
                '', # 비고
                ''  # (Full작업자) hidden storage
            )
            self.monthly_usage_tree.insert('', tk.END, values=total_values, tags=('total',))
            
            # --- [NEW] Populate Summaries (Initial/Total) ---
            self._populate_monthly_summary_trees(df, has_note)
            
            # --- Dynamic Column Hiding ---
            mandatory_cols = ['연도', '월', '현장', '작업자', '작업시간', '품목명']
            dynamic_col_status = {
                'OT시간': total_ot_hours > 0,
                'OT금액': total_ot_amount > 0,
                '수량': total_test_amount > 0,
                '단가': total_unit_price > 0,
                '출장비': total_travel_cost > 0,
                '일식': total_meal_cost > 0,
                '검사비': total_test_fee > 0,
                '필름매수': total_film_count > 0,
                '센터미스': total_rtk_center > 0,
                '농도': total_rtk_density > 0,
                '마킹미스': total_rtk_marking > 0,
                '필름마크': total_rtk_film > 0,
                '취급부주의': total_rtk_handling > 0,
                '고객불만': total_rtk_customer > 0,
                '기타': total_rtk_other > 0,
                'RTK총계': total_rtk_sum > 0,
                '형광자분': total_ndt_fluorescent_mag > 0,
                '흑색자분': total_ndt_magnet > 0,
                '백색페인트': total_ndt_paint > 0,
                '침투제': total_ndt_penetrant > 0,
                '세척제': total_ndt_cleaner > 0,
                '현상제': total_ndt_developer > 0,
                '형광침투제': total_ndt_fluorescent_pen > 0,
                '비고': has_note
            }
            # OT1-10 status
            for i in range(1, 11):
                dynamic_col_status[f'OT{i}'] = total_indiv_ot_amounts[i-1] > 0
                
            all_cols = self.monthly_usage_tree['columns']
            visible_cols = [col for col in all_cols if col in mandatory_cols or dynamic_col_status.get(col, False)]
            self.monthly_usage_tree['displaycolumns'] = visible_cols
            
            # Ensure proper configuration for visible columns
            for col in visible_cols:
                self.monthly_usage_tree.column(col, stretch=False, minwidth=20)
            
            # Re-apply saved widths
            if hasattr(self, 'tab_config') and 'monthly_usage_col_widths' in self.tab_config:
                saved_widths = self.tab_config['monthly_usage_col_widths']
                for col in visible_cols:
                    if col in saved_widths:
                        try: self.monthly_usage_tree.column(col, width=int(saved_widths[col]))
                        except: pass

            # Ensure Total Row stays at bottom
            self.monthly_usage_tree.detach(self.monthly_usage_tree.get_children()[-1])
            self.monthly_usage_tree.insert('', tk.END, values=total_values, tags=('total',))
        else:
            # If empty data
            visible_cols = ['연도', '월', '현장', '작업자', '작업시간', '품목명']
            self.monthly_usage_tree['displaycolumns'] = visible_cols
            for col in visible_cols:
                self.monthly_usage_tree.column(col, stretch=False, minwidth=20)

    def on_monthly_usage_select(self, event):
        """Update site and worker summaries when a row is selected in monthly usage tree"""
        selection = self.monthly_usage_tree.selection()
        if not selection:
            return
            
        item = selection[0]
        values = self.monthly_usage_tree.item(item, 'values')
        tags = self.monthly_usage_tree.item(item, 'tags')
        
        # If no data stored yet or error
        if not hasattr(self, 'current_monthly_df') or self.current_monthly_df.empty:
            return
            
        # If total row is selected, show total summaries
        if 'total' in tags:
            self._populate_monthly_summary_trees(self.current_monthly_df)
            return
            
        # Extract row info (Year, Month, Site, Material)
        try:
            year = int(values[0])
            month = int(values[1])
            site = str(values[2]).strip()
            # Material is at index 22 (after 연도-월-현장-작업자-작업시간-OT시간-OT금액-OT1...OT10-검사량-단가-출장비-일식-검사비)
            mat_name = str(values[22]).strip()
            
            # Filter the current monthly dataset
            mask = (self.current_monthly_df['Year'] == year) & \
                   (self.current_monthly_df['Month'] == month) & \
                   (self.current_monthly_df['Site'] == site)
            
            # For material, find matching ID
            matching_rows = []
            matching_ids = []
            for m_id, m_name in zip(self.materials_df['MaterialID'], self.materials_df['품목명']):
                if m_name == mat_name:
                    matching_ids.append(m_id)
            
            if matching_ids:
                mask = mask & (self.current_monthly_df['MaterialID'].isin(matching_ids))
            
            filtered_subset = self.current_monthly_df[mask]
            self._populate_monthly_summary_trees(filtered_subset)
            
        except Exception as e:
            print(f"DEBUG: Error in on_monthly_usage_select: {e}")

    def _populate_monthly_summary_trees(self, df, has_note=None):
        """Helper to fill site and worker summary trees with given data subset"""
        # Clear current views
        for item in self.site_summary_tree.get_children():
            self.site_summary_tree.delete(item)
        for item in self.worker_summary_tree.get_children():
            self.worker_summary_tree.delete(item)
            
        if df.empty:
            return

        # --- Populate Site Summary ---
        # Include MaterialID in grouping and add relevant NDT/Cost fields to aggregation
        site_agg_dict = {
            '검사비': 'sum',
            '출장비': 'sum',
            'FilmCount': 'sum'
        }
        
        # Add RTK fields for RT매수 calculation
        rtk_fields = ['RTK_센터미스', 'RTK_농도', 'RTK_마킹미스', 'RTK_필름마크', 'RTK_취급부주의', 'RTK_고객불만', 'RTK_기타']
        for rf in rtk_fields:
            if rf in df.columns: site_agg_dict[rf] = 'sum'
            
        # Add NDT fields
        ndt_map = {
            'NDT_형광자분': 'sum',
            'NDT_자분': 'sum',
            'NDT_흑색자분': 'sum',
            'NDT_페인트': 'sum',
            'NDT_백색페인트': 'sum',
            'NDT_침투제': 'sum',
            'NDT_세척제': 'sum',
            'NDT_현상제': 'sum',
            'NDT_형광': 'sum',
            'NDT_형광침투제': 'sum'
        }
        for nf in ndt_map:
            if nf in df.columns: site_agg_dict[nf] = 'sum'

        site_summary = df.groupby(['Site', 'MaterialID']).agg(site_agg_dict).reset_index()
        
        for _, row in site_summary.iterrows():
            # Get FilmCount (now labeled as 필름매수)
            film_total = row.get('FilmCount', 0.0)
            
            # Calculate 흑색자분 (NDT_자분 + NDT_흑색자분)
            black_mag = row.get('NDT_자분', 0.0) + row.get('NDT_흑색자분', 0.0)
            
            # Calculate 백색페인트 (NDT_페인트 + NDT_백색페인트)
            white_paint = row.get('NDT_페인트', 0.0) + row.get('NDT_백색페인트', 0.0)
            
            # PT materials
            pt_pen = row.get('NDT_침투제', 0.0)
            pt_cln = row.get('NDT_세척제', 0.0)
            pt_dev = row.get('NDT_현상제', 0.0)
            pt_fluoro = row.get('NDT_형광', 0.0) + row.get('NDT_형광침투제', 0.0)
            
            # RTK details and total (센터미스 ~ 기타 합계)
            r_center = row.get('RTK_센터미스', 0.0)
            r_density = row.get('RTK_농도', 0.0)
            r_marking = row.get('RTK_마킹미스', 0.0)
            r_film = row.get('RTK_필름마크', 0.0)
            r_careless = row.get('RTK_취급부주의', 0.0)
            r_customer = row.get('RTK_고객불만', 0.0)
            r_other = row.get('RTK_기타', 0.0)
            rtk_sum = r_center + r_density + r_marking + r_film + r_careless + r_customer + r_other
            
            # Get Material Name
            mat_id = row['MaterialID']
            mat_row = self.materials_df[self.materials_df['MaterialID'] == mat_id]
            mat_name = mat_row.iloc[0]['품목명'] if not mat_row.empty else f"ID: {mat_id}"

            self.site_summary_tree.insert('', tk.END, values=(
                row['Site'],
                mat_name,
                f"{film_total:.1f}" if film_total > 0 else '',
                f"{row['검사비']:,.0f}" if row['검사비'] > 0 else '',
                f"{row['출장비']:,.0f}" if row['출장비'] > 0 else '',
                f"{row.get('NDT_형광자분', 0.0):.1f}" if row.get('NDT_형광자분', 0.0) > 0 else '',
                f"{black_mag:.1f}" if black_mag > 0 else '',
                f"{white_paint:.1f}" if white_paint > 0 else '',
                f"{pt_pen:.1f}" if pt_pen > 0 else '',
                f"{pt_cln:.1f}" if pt_cln > 0 else '',
                f"{pt_dev:.1f}" if pt_dev > 0 else '',
                f"{pt_fluoro:.1f}" if pt_fluoro > 0 else '',
                f"{r_center:.1f}" if r_center > 0 else '',
                f"{r_density:.1f}" if r_density > 0 else '',
                f"{r_marking:.1f}" if r_marking > 0 else '',
                f"{r_film:.1f}" if r_film > 0 else '',
                f"{r_careless:.1f}" if r_careless > 0 else '',
                f"{r_customer:.1f}" if r_customer > 0 else '',
                f"{r_other:.1f}" if r_other > 0 else '',
                f"{rtk_sum:.1f}" if rtk_sum > 0 else ''
            ))

        # --- Populate Worker Summary ---
        worker_data = []
        for i in range(1, 11):
            u_c = 'User' if i == 1 else f'User{i}'
            w_c = 'WorkTime' if i == 1 else f'WorkTime{i}'
            ot_c = 'OT' if i == 1 else f'OT{i}'
            if u_c in df.columns and w_c in df.columns:
                temp_df = df[[u_c, w_c, ot_c]].copy()
                temp_df.columns = ['WorkerName', 'ShiftType', 'OTValue']
                worker_data.append(temp_df)
        
        if worker_data:
            worker_df = pd.concat(worker_data)
            worker_df['WorkerName'] = worker_df['WorkerName'].apply(self.clean_nan)
            worker_df = worker_df[worker_df['WorkerName'] != '']
            
            if not worker_df.empty:
                # [NEW] Enhanced OT hour extraction: check WorkTime if OTValue is only amount
                def get_ot_hours_robust(row):
                    ot_val = str(row['OTValue']).strip()
                    if not ot_val or ot_val == '0': return 0.0
                    
                    # If it's already "N시간...", parse it
                    if '시간' in ot_val:
                        return self._parse_ot_hours(ot_val)
                    
                    # If it's just an amount, calculate hours from WorkTime
                    if ot_val.replace(',', '').isdigit() and int(ot_val.replace(',', '')) > 100:
                        wt_val = row['ShiftType']
                        date_val = pd.to_datetime(df.iloc[0]['Date']) # Approximation for subset date context
                        h, _ = self._calculate_ot_from_worktime(wt_val, date_val)
                        return h
                    
                    return self._parse_ot_hours(ot_val)

                worker_df['OT_H'] = worker_df.apply(get_ot_hours_robust, axis=1)
                
                # Split calculation also needs to be robust
                def get_split_robust(row):
                    ot_val = str(row['OTValue']).strip()
                    date_val = pd.to_datetime(df.iloc[0]['Date'])
                    if '시간' in ot_val or not (ot_val.replace(',', '').isdigit() and int(ot_val.replace(',', '')) > 100):
                        return self._calculate_split_ot_hours(ot_val, date_val)
                    else:
                        # Recalculate from WorkTime
                        wt_val = row['ShiftType']
                        
                        # We need the simulation logic from _calculate_split_ot_hours but using wt_val
                        # Let's use a temporary string that _calculate_split_ot_hours can parse
                        h, _ = self._calculate_ot_from_worktime(wt_val, date_val)
                        fake_ot_str = f"{h}시간 ({ot_val}원)"
                        # Check start hour from wt_val for range logic
                        marker_pattern = MARKER_PATTERN
                        clean_wt = marker_pattern.sub('', str(wt_val)).strip()
                        if '~' in clean_wt:
                            start_time = clean_wt.split('~')[0]
                            fake_ot_str = f"{start_time}~{h}시간 ({ot_val}원)"
                        
                        return self._calculate_split_ot_hours(fake_ot_str, date_val)

                # Now returns 3 values (day, night, holiday_dawn)
                worker_df[['OT_H_Day', 'OT_H_Night', 'OT_H_Holiday_Dawn']] = worker_df.apply(
                    lambda r: pd.Series(get_split_robust(r)), axis=1
                )
                worker_df['OT_A'] = worker_df['OTValue'].apply(self.calculate_ot_amount)
                worker_df['Count'] = 1.0
                
                # Shift types (explicit string markers)
                worker_df['Shift_Holiday'] = worker_df['ShiftType'].apply(lambda x: 1.0 if '휴일' in str(x) else 0.0)
                worker_df['Shift_Day'] = worker_df['ShiftType'].apply(lambda x: 1.0 if '주간' in str(x) and '휴일' not in str(x) else 0.0)
                worker_df['Shift_Night'] = worker_df['ShiftType'].apply(lambda x: 1.0 if '야간' in str(x) and '휴일' not in str(x) else 0.0)
                
                # If explicit holiday shift, all OT is holiday.
                # If NOT explicit holiday shift, add the OT_H_Holiday_Dawn (Friday dawn) to holiday hours.
                worker_df['H_Holiday'] = worker_df.apply(lambda r: r['OT_H'] if r['Shift_Holiday'] > 0 else r['OT_H_Holiday_Dawn'], axis=1)
                
                # Calculate amount properly. If it was fully holiday, it's just OT_A.
                # If it was Friday Dawn, add the calculated dawn hours * 7500 to whatever other OT they had?
                # Actually, A_Holiday should just be the amount from those hours
                worker_df['A_Holiday'] = worker_df.apply(lambda r: r['OT_A'] if r['Shift_Holiday'] > 0 else r['OT_H_Holiday_Dawn'] * 7500, axis=1)
                
                # For day/night, they only get values if it's NOT a full holiday shift.
                worker_df['H_Day'] = worker_df.apply(lambda r: r['OT_H_Day'] if r['Shift_Holiday'] == 0 else 0.0, axis=1)
                worker_df['H_Night'] = worker_df.apply(lambda r: r['OT_H_Night'] if r['Shift_Holiday'] == 0 else 0.0, axis=1)
                worker_df['A_Day'] = worker_df.apply(lambda r: r['H_Day'] * 4000 if r['Shift_Holiday'] == 0 else 0.0, axis=1)
                worker_df['A_Night'] = worker_df.apply(lambda r: r['H_Night'] * 5000 if r['Shift_Holiday'] == 0 else 0.0, axis=1)

                worker_summary = worker_df.groupby('WorkerName').agg({
                    'Count': 'sum',
                    'H_Day': 'sum',
                    'H_Night': 'sum',
                    'H_Holiday': 'sum',
                    'OT_H': 'sum',
                    'A_Day': 'sum',
                    'A_Night': 'sum',
                    'A_Holiday': 'sum',
                    'OT_A': 'sum'
                }).reset_index()
                
                for _, row in worker_summary.iterrows():
                    self.worker_summary_tree.insert('', tk.END, values=(
                        row['WorkerName'],
                        f"{row['Count']:.1f}",
                        f"{row['H_Day']:.1f}",
                        f"{row['H_Night']:.1f}",
                        f"{row['H_Holiday']:.1f}",
                        f"{row['OT_H']:.1f}",
                        f"{row['A_Day']:,.0f}",
                        f"{row['A_Night']:,.0f}",
                        f"{row['A_Holiday']:,.0f}",
                        f"{row['OT_A']:,.0f}"
                    ))

    def apply_worker_shift_hours_to_budget(self):
        """월별 탭의 작업자별 누계(주간/야간/휴일) 시간을 공사실행예산서 특별근무 투입시간에 적용한다."""
        if not hasattr(self, 'labor_detail_widget'):
            messagebox.showwarning("공사탭 미초기화",
                                   "공사실행예산서 탭을 먼저 열어 초기화하세요.")
            return

        # worker_summary_tree에서 모든 행의 주간/야간/휴일 합산
        total_h_day = 0.0
        total_h_night = 0.0
        total_h_holiday = 0.0
        worker_count_day = 0
        worker_count_night = 0
        worker_count_holiday = 0

        def _f(val):
            try: return float(str(val).replace(',', '') or 0)
            except: return 0.0

        for item in self.worker_summary_tree.get_children():
            vals = self.worker_summary_tree.item(item, 'values')
            # 컬럼 순서: 작업자, 총공수, 연장(시간), 야간(시간), 휴일(시간), 총OT(시간), 연장(금액), 야간(금액), 휴일(금액), 총OT(금액)
            h_day     = _f(vals[2]) if len(vals) > 2 else 0.0
            h_night   = _f(vals[3]) if len(vals) > 3 else 0.0
            h_holiday = _f(vals[4]) if len(vals) > 4 else 0.0
            if h_day > 0:
                total_h_day += h_day
                worker_count_day += 1
            if h_night > 0:
                total_h_night += h_night
                worker_count_night += 1
            if h_holiday > 0:
                total_h_holiday += h_holiday
                worker_count_holiday += 1

        if total_h_day == 0 and total_h_night == 0 and total_h_holiday == 0:
            messagebox.showinfo("데이터 없음",
                                "작업자별 누계에 주간/야간/휴일 시간 데이터가 없습니다.\n"
                                "월별 탭에서 조회를 먼저 실행하세요.")
            return

        # LaborCostDetailWidget 특별근무 섹션에 입력
        ldw = self.labor_detail_widget
        # 연장근무 ← 주간 OT 시간 합계 / 인원수
        if worker_count_day > 0:
            avg_day = total_h_day / worker_count_day
            ldw.entries["연장근무"]['personnel'].delete(0, 'end')
            ldw.entries["연장근무"]['personnel'].insert(0, f"{worker_count_day:g}")
            ldw.entries["연장근무"]['period'].delete(0, 'end')
            ldw.entries["연장근무"]['period'].insert(0, f"{avg_day:.1f}")
        # 야간근무 ← 야간 OT 시간 합계
        if worker_count_night > 0:
            avg_night = total_h_night / worker_count_night
            ldw.entries["야간근무"]['personnel'].delete(0, 'end')
            ldw.entries["야간근무"]['personnel'].insert(0, f"{worker_count_night:g}")
            ldw.entries["야간근무"]['period'].delete(0, 'end')
            ldw.entries["야간근무"]['period'].insert(0, f"{avg_night:.1f}")
        # 휴일근무 ← 휴일 OT 시간 합계
        if worker_count_holiday > 0:
            avg_holiday = total_h_holiday / worker_count_holiday
            ldw.entries["휴일근무"]['personnel'].delete(0, 'end')
            ldw.entries["휴일근무"]['personnel'].insert(0, f"{worker_count_holiday:g}")
            ldw.entries["휴일근무"]['period'].delete(0, 'end')
            ldw.entries["휴일근무"]['period'].insert(0, f"{avg_holiday:.1f}")

        # 계산 반영
        ldw.calculate_all()

        # 공사실행예산서 탭으로 이동
        try:
            tab_idx = [self.notebook.tab(i, 'text') for i in range(self.notebook.index('end'))].index('공사실행예산서')
            self.notebook.select(tab_idx)
        except:
            pass

        messagebox.showinfo("적용 완료",
                            f"특별근무 투입시간 적용 완료!\n"
                            f"  연장근무: {worker_count_day}명 / {total_h_day:.1f}h (평균 {total_h_day/max(worker_count_day,1):.1f}h/인)\n"
                            f"  야간근무: {worker_count_night}명 / {total_h_night:.1f}h (평균 {total_h_night/max(worker_count_night,1):.1f}h/인)\n"
                            f"  휴일근무: {worker_count_holiday}명 / {total_h_holiday:.1f}h (평균 {total_h_holiday/max(worker_count_holiday,1):.1f}h/인)")

    def export_monthly_usage_history(self):
        """Export monthly usage data, site summaries, and worker summaries to a multi-sheet Excel file"""
        try:
            # 1. Collect Main Monthly Usage Data
            columns = self.monthly_usage_tree['columns']
            monthly_data = []
            for item in self.monthly_usage_tree.get_children():
                monthly_data.append(self.monthly_usage_tree.item(item, 'values'))
            
            if not monthly_data:
                messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
                return

            # 2. Collect Site Summary Data
            site_columns = self.site_summary_tree['columns']
            site_data = []
            for item in self.site_summary_tree.get_children():
                site_data.append(self.site_summary_tree.item(item, 'values'))

            # 3. Collect Worker Summary Data
            worker_columns = self.worker_summary_tree['columns']
            worker_data = []
            for item in self.worker_summary_tree.get_children():
                worker_data.append(self.worker_summary_tree.item(item, 'values'))
            
            # Prepare filename
            today = datetime.datetime.now().strftime('%Y%m%d')
            # Using a descriptive filename indicating it's a comprehensive report (종합)
            filename = f"월별집계_종합_{today}.xlsx"
            
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=filename,
                title="월별 집계(종합) 내역 저장",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not save_path:
                return

            # --- Process Monthly Usage Tree (Sheet 1) ---
            columns_to_export = [c for c in columns if c != '(Full작업자)']
            col_to_idx = {col: i for i, col in enumerate(columns)}
            
            final_monthly_data = []
            for row in monthly_data:
                new_row = list(row)
                # Swap summarized worker with full list if (Full작업자) exists
                if '작업자' in col_to_idx and '(Full작업자)' in col_to_idx:
                    full_idx = col_to_idx['(Full작업자)']
                    if full_idx < len(new_row):
                        full_val = new_row[full_idx]
                        if full_val:
                            new_row[col_to_idx['작업자']] = full_val
                
                # Filter only display columns
                final_row = [new_row[i] for i, c in enumerate(columns) if c != '(Full작업자)' and i < len(new_row)]
                final_monthly_data.append(final_row)
            
            monthly_df = pd.DataFrame(final_monthly_data, columns=columns_to_export)
            monthly_df = self.clean_df_export(monthly_df)

            # --- Process Site Summary (Sheet 2) ---
            site_df = pd.DataFrame(site_data, columns=site_columns)
            site_df = self.clean_df_export(site_df)

            # --- Process Worker Summary (Sheet 3) ---
            worker_df = pd.DataFrame(worker_data, columns=worker_columns)
            worker_df = self.clean_df_export(worker_df)

            # --- Save to Excel with Multiple Sheets and AutoFit ---
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                # Local helper for autofitting to avoid polluting global namespace
                def write_sheet(df, sheet_name):
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    worksheet = writer.sheets[sheet_name]
                    
                    def get_display_width(s):
                        width = 0
                        for char in str(s):
                            if ord(char) > 127: width += 2 # Double width for Korean
                            else: width += 1
                        return width

                    for idx, col in enumerate(df.columns):
                        series = df[col].astype(str)
                        lengths = series.apply(get_display_width)
                        max_val_len = lengths.max() if not lengths.empty else 0
                        header_len = get_display_width(col)
                        # Cap at 40 to allow better horizontal density in summaries
                        final_width = min(max(max_val_len, header_len) + 2, 40)
                        
                        col_letter = worksheet.cell(row=1, column=idx+1).column_letter
                        worksheet.column_dimensions[col_letter].width = final_width

                write_sheet(monthly_df, "월별집계_내역")
                write_sheet(site_df, "현장별_누계")
                write_sheet(worker_df, "작업자별_누계")

            messagebox.showinfo("완료", "월별 집계 내역(종합)이 저장되었습니다.\n(내역, 현장별, 작업자별 3개 시트 포함)")
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패: {e}")
            traceback.print_exc()

    def create_draggable_container(self, parent, label_text, widget_class, config_key, manage_list_key=None, grid_info=None, **widget_kwargs):
        """Create a draggable container with a label and a widget, styled as a visible box"""
        # Use ttk.Frame for automatic background matching
        container = ttk.Frame(parent, relief="solid", borderwidth=1)
        
        # Header container for label and buttons
        hdr = ttk.Frame(container) 
        hdr.pack(side='top', fill='x', padx=0, pady=0)
        
        # Drag handle icon
        lbl_drag = ttk.Label(hdr, text="✥", font=('Arial', 9), cursor='fleur')
        lbl_drag.pack(side='left', padx=1)
        self.make_header_draggable(lbl_drag, container)
        
        # Label
        lbl = ttk.Label(hdr, text=label_text, font=('Malgun Gothic', 9, 'bold'))
        lbl.pack(side='left', padx=1)
        self.make_header_draggable(lbl, container)
        
        # Icons container on the right
        btn_box = ttk.Frame(hdr)
        btn_box.pack(side='right', padx=1)

        # Rename icon
        btn_rename = ttk.Label(btn_box, text="✏️", font=('Arial', 8), cursor='hand2')
        btn_rename.pack(side='left', padx=1)
        btn_rename.bind('<Button-1>', lambda e: self.rename_widget_label(config_key))
        
        # Clone icon
        btn_clone = ttk.Label(btn_box, text="📋", font=('Arial', 8), cursor='hand2')
        btn_clone.pack(side='left', padx=1)
        btn_clone.bind('<Button-1>', lambda e: self.clone_widget(config_key))

        # Delete icon (X)
        btn_del = ttk.Label(btn_box, text="❌", font=('Arial', 8), cursor='hand2')
        btn_del.pack(side='left', padx=1)
        btn_del.bind('<Button-1>', lambda e: self.remove_box(config_key))

        # Manage List Icon (Gear) - if it's a list-based widget
        if manage_list_key:
            btn_manage = ttk.Label(btn_box, text="⚙️", font=('Arial', 8), cursor='hand2')
            btn_manage.pack(side='left', padx=1)
            btn_manage.bind('<Button-1>', lambda e: self.open_list_management_dialog(manage_list_key))
        
        # Internal Content Area
        content_area = ttk.Frame(container, padding=(1, 0))
        content_area.pack(side='top', fill='both', expand=True)

        # Widget
        widget = widget_class(content_area, **widget_kwargs)
        widget.pack(side='left', fill='both', expand=True)
        
        # If the widget is a basic tk widget (Text, Canvas), set its background
        if hasattr(widget, 'config') and 'bg' in widget.keys():
            try: widget.config(bg=self.theme_bg)
            except: pass
        
        # NEW: Handle grid placement first if grid_info provided
        if grid_info:
            container.grid(**grid_info)
        
        # Track for layout reset/config
        container._config_key = config_key
        container._label_widget = lbl
        container._widget = widget
        container._widget_class = widget_class
        container._widget_kwargs = widget_kwargs
        container._manage_list_key = manage_list_key
        
        # Register and make draggable (this now captures the CORRECT grid info)
        self.draggable_items[config_key] = container
        self.make_draggable(container, config_key)
        
        return container, widget

    def open_list_management_dialog(self, config_key):
        """Open a generic dialog to manage (edit/delete) items in a data list"""
        if self.layout_locked: return

        # ── 이미 열린 창이 있으면 앞으로 가져오고 종료 ──────────────
        if not hasattr(self, '_list_mgmt_dialogs'):
            self._list_mgmt_dialogs = {}
        existing = self._list_mgmt_dialogs.get(config_key)
        if existing and existing.winfo_exists():
            existing.lift()
            existing.focus_set()
            return

        # Map key to actual list
        data_map = {
            'sites': ('현장 목록 관리', self.sites),
            'users': ('담당자 목록 관리', getattr(self, 'users', [])),
            'equipments': ('장비 목록 관리', getattr(self, 'equipments', [])),
            'vehicles': ('차량 목록 관리', getattr(self, 'vehicles', []))
        }
        
        if config_key not in data_map: return
        title, data_list = data_map[config_key]
        
        dialog = tk.Toplevel(self.root)
        self._list_mgmt_dialogs[config_key] = dialog  # 창 추적 등록
        dialog.title(title)
        dialog.geometry("400x400")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding=10)
        frame.pack(fill='both', expand=True)
        
        listbox = tk.Listbox(frame, font=('Arial', 10))
        listbox.pack(fill='both', expand=True, side='left')
        
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=listbox.yview)
        scrollbar.pack(fill='y', side='right')
        listbox.config(yscrollcommand=scrollbar.set)
        
        for item in sorted(data_list):
            listbox.insert('end', item)
            
        btn_frame = ttk.Frame(dialog, padding=5)
        btn_frame.pack(fill='x')
        
        def refresh_list():
            listbox.delete(0, 'end')
            for item in sorted(data_list):
                listbox.insert('end', item)
            # Trigger app-wide update of related comboboxes
            self.refresh_ui_for_list_change(config_key)

        def add_item():
            new_val = simpledialog.askstring("추가", "새 이름을 입력하세요:")
            if new_val and new_val.strip():
                val = new_val.strip()
                if val not in data_list:
                    data_list.append(val)
                    data_list.sort()
                    self.save_tab_config()
                    refresh_list()
                else:
                    messagebox.showinfo("정보", "이미 목록에 있는 이름입니다.")

        def edit_item():
            sel = listbox.curselection()
            if not sel: return
            idx = sel[0]
            old_val = listbox.get(idx)
            new_val = simpledialog.askstring("수정", "새 이름을 입력하세요:", initialvalue=old_val)
            if new_val and new_val.strip() and new_val != old_val:
                if old_val in data_list: data_list.remove(old_val)
                data_list.append(new_val.strip())
                data_list.sort()
                self.save_tab_config()
                refresh_list()

        def delete_item():
            sel = listbox.curselection()
            if not sel: return
            idx = sel[0]
            val = listbox.get(idx)
            if messagebox.askyesno("삭제 확인", f"'{val}'을 목록에서 삭제하시겠습니까?"):
                if val in data_list: data_list.remove(val)
                self.save_tab_config()
                refresh_list()

        def on_close():
            self._list_mgmt_dialogs.pop(config_key, None)
            dialog.destroy()

        dialog.protocol("WM_DELETE_WINDOW", on_close)  # X 버튼도 추적 해제

        ttk.Button(btn_frame, text="추가", command=add_item).pack(side='left', padx=5, expand=True)
        ttk.Button(btn_frame, text="수정", command=edit_item).pack(side='left', padx=5, expand=True)
        ttk.Button(btn_frame, text="삭제", command=delete_item).pack(side='left', padx=5, expand=True)
        ttk.Button(btn_frame, text="닫기", command=on_close).pack(side='left', padx=5, expand=True)

    def refresh_ui_for_list_change(self, config_key):
        """Update all related UI elements after a list (sites, users, etc) has changed"""
        # Dictionary mapping config keys to their current values
        list_map = {
            'sites': self.sites,
            'users': self.users,
            'warehouses': self.warehouses,
            'equipments': self.equipments,
            'worktimes': self.worktimes,
            'vehicles': self.vehicles
        }
        
        if config_key not in list_map:
            return
            
        current_vals = list_map[config_key]
        sorted_vals = sorted(current_vals)
        
        # 1. Update standard widgets
        if config_key == 'sites':
            if hasattr(self, 'cb_daily_site'): self.cb_daily_site['values'] = sorted_vals
            if hasattr(self, 'cb_trans_site'): self.cb_trans_site['values'] = sorted_vals
            if hasattr(self, 'cb_daily_filter_site'):
                self.cb_daily_filter_site['values'] = ['전체'] + sorted_vals
            if hasattr(self, 'cb_budget_site'): self.cb_budget_site['values'] = sorted_vals
            if hasattr(self, 'cb_budget_view_site'): self.cb_budget_view_site['values'] = sorted_vals
        elif config_key == 'users':
            # Updated to match current attribute names if needed, 
            # but usually it's cb_daily_user for 1, and cb_daily_user{i} for 2-10
            for i in range(1, 11):
                attr = 'cb_daily_user' if i == 1 else f'cb_daily_user{i}'
                if hasattr(self, attr):
                    widget = getattr(self, attr)
                    if isinstance(widget, WorkerCompositeWidget):
                         widget.cb_name['values'] = sorted_vals
                    elif hasattr(widget, 'configure'):
                        try: widget['values'] = sorted_vals
                        except: pass
            
            if hasattr(self, 'cb_trans_user'): self.cb_trans_user['values'] = sorted_vals
        elif config_key == 'worktimes':
            # Update all worktime fields (1-10)
            for i in range(1, 11):
                attr = f'ent_worktime{i}'
                if hasattr(self, attr):
                    widget = getattr(self, attr)
                    if hasattr(widget, 'configure'):
                        try: 
                            widget['values'] = sorted_vals
                        except: pass
        elif config_key == 'equipments':
            self.update_material_combo()
        elif config_key == 'vehicles':
            # Update all VehicleInspectionWidgets
            for key, widget_instance in self.vehicle_inspections.items():
                widget_instance.update_vehicle_list(sorted_vals)


        # 2. Update ALL draggable widgets (clones) that depend on this list
        for key, container in self.draggable_items.items():
            # Heuristic: if manage_list_key is missing but label suggests it's a worker/user/site
            m_key = getattr(container, '_manage_list_key', None)
            if not m_key:
                if hasattr(container, '_label_widget'):
                    lbl_text = container._label_widget.cget('text').lower()
                    if config_key == 'users' and any(x in lbl_text for x in ['작업자', '담당자', 'user', 'worker']):
                        m_key = 'users'
                        container._manage_list_key = 'users'
                    elif config_key == 'sites' and any(x in lbl_text for x in ['현장', 'site']):
                        m_key = 'sites'
                        container._manage_list_key = 'sites'
                    elif config_key == 'equipments' and any(x in lbl_text for x in ['장비', 'equip']):
                        m_key = 'equipments'
                        container._manage_list_key = 'equipments'
                    elif config_key == 'vehicles' and any(x in lbl_text for x in ['차량', 'vehicle']):
                        m_key = 'vehicles'
                        container._manage_list_key = 'vehicles'

            if m_key == config_key:
                if hasattr(container, '_widget') and hasattr(container._widget, 'config'):
                    try:
                        container._widget['values'] = sorted_vals
                    except:
                        pass

    def remove_box(self, key):
        """Intelligently remove a box: hide standard ones, destroy custom ones"""
        is_dynamic = key.startswith('memo_') or key.startswith('clone_') or key.startswith('checklist_') or key.startswith('vehicle_inspection_')
        
        # Allow deletion of dynamic widgets even if layout is locked
        if self.layout_locked and not is_dynamic: return
        
        if is_dynamic:
            # Permanent deletion for custom/dynamic items
            self.destroy_custom_widget(key)
        else:
            # Hiding for standard items (can be restored via Reset All)
            widget = self.draggable_items.get(key)
            if widget:
                # We reuse the existing hide_widget logic
                self.hide_widget(None, widget=widget)

    def destroy_custom_widget(self, key):
        """Destroy and remove from config any dynamic widget (clone or memo)"""
        # Allow destruction even if layout_locked (since it's a content management action)
        
        widget = self.draggable_items.get(key)
        if widget:
            widget.destroy()
            if key in self.draggable_items:
                del self.draggable_items[key]
            if key in self.memos:
                del self.memos[key]
            if key in self.checklists:
                del self.checklists[key]
            if key in self.vehicle_inspections:
                del self.vehicle_inspections[key]
            
            # Clean from in-memory config to prevent resurrection on next save_tab_config call
            if hasattr(self, 'tab_config') and 'draggable_geometries' in self.tab_config:
                if key in self.tab_config['draggable_geometries']:
                    del self.tab_config['draggable_geometries'][key]
            
            # Clean from config immediately to prevent resurrection on next load
            try:
                import json
                if os.path.exists(self.config_path):
                    with open(self.config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                    
                    if 'draggable_geometries' in config and key in config['draggable_geometries']:
                        del config['draggable_geometries'][key]
                        
                    with open(self.config_path, 'w', encoding='utf-8') as f:
                        json.dump(config, f, ensure_ascii=False, indent=2)
            except:
                pass
            
            self.save_tab_config()

    def rename_widget_label(self, key):
        """Show a simple dialog to rename a widget's label"""
        if self.layout_locked: return
        
        widget = self.draggable_items.get(key)
        if not widget: return
        
        current_text = ""
        if hasattr(widget, '_label_widget'):
            current_text = widget._label_widget.cget('text')
        elif key in self.memos:
            current_text = self.memos[key]['title_entry'].get()
        elif key in self.checklists:
            current_text = self.checklists[key]['title_entry'].get()
            
        new_name = simpledialog.askstring("이름 변경", "새 이름을 입력하세요:", initialvalue=current_text)
        if new_name is not None:
            if hasattr(widget, '_label_widget'):
                widget._label_widget.config(text=new_name)
            elif key in self.memos:
                self.memos[key]['title_entry'].delete(0, 'end')
                self.memos[key]['title_entry'].insert(0, new_name)
            elif key in self.checklists:
                self.checklists[key]['title_entry'].delete(0, 'end')
                self.checklists[key]['title_entry'].insert(0, new_name)
            self.save_tab_config()

    def clone_widget(self, key):
        """Create a clone of an existing widget as a new custom box"""
        if self.layout_locked: return
        
        orig = self.draggable_items.get(key)
        if not orig: return
        
        import time
        new_key = f"clone_{int(time.time() * 1000)}"
        
        label_text = ""
        if hasattr(orig, '_label_widget'):
            label_text = orig._label_widget.cget('text')
        
        if hasattr(orig, '_widget_class'):
            # It's a container created via create_draggable_container
            cont, w = self.create_draggable_container(
                self.entry_inner_frame, 
                label_text, 
                orig._widget_class, 
                new_key, 
                manage_list_key=getattr(orig, '_manage_list_key', None), # Pass manage_list_key
                **orig._widget_kwargs
            )
            
            # Copy value from original widget
            if hasattr(orig, '_widget'):
                try:
                    current_val = str(orig._widget.get()) # Ensure string
                    
                    # Try generic Entry-like setting (works for Entry and Combobox text area)
                    if hasattr(w, 'delete') and hasattr(w, 'insert'):
                        try:
                            w.delete(0, 'end')
                            w.insert(0, current_val)
                        except:
                            # Readonly comboboxes might fail delete/insert
                            pass
                            
                    # Try specific set method (Combobox, Scale, etc)
                    if hasattr(w, 'set'):
                        w.set(current_val)
                        
                except Exception as e:
                    print(f"Failed to copy value: {e}")
            
            cont.place(x=50, y=50) # Start position
            self.save_tab_config()

        elif key in self.memos:
            # It's a memo
            content = self.memos[key]['text_widget'].get('1.0', 'end-1c')
            title = self.memos[key]['title_entry'].get()
            self.add_new_memo(initial_text=content, initial_title=title, key=new_key)
            self.save_tab_config()
        elif key in self.checklists:
            # It's a checklist
            self.duplicate_checklist(key)
            self.save_tab_config()

    def _bind_recursive(self, widget, target_container):
        """Recursively bind drag events to widget and its children"""
        from functools import partial
        
        # Bind events to this widget, targeting the container
        # Note: We use add=True to avoid overwriting existing bindings if possible, 
        # but for drag we usually want exclusive control or at least priority.
        # Here we just bind standard.
        
        # We need to capture the target_container in the callback
        widget.bind("<Button-3>", partial(self.on_drag_start, widget=target_container))
        widget.bind("<Shift-Button-3>", partial(self.on_resize_start, widget=target_container))
        widget.bind("<B3-Motion>", partial(self.on_mouse_motion, widget=target_container))
        widget.bind("<Double-Button-3>", partial(self.reset_widget_position, widget=target_container))
        widget.bind("<Control-Button-3>", partial(self.hide_widget, widget=target_container))
        widget.bind("<ButtonRelease-3>", partial(self.on_drag_stop, widget=target_container))
        
        # Recurse for children
        try:
            for child in widget.winfo_children():
                # Toplevel(DateEntry 달력 팝업 등) 내부는 절대 순회하지 않음
                # → winfo_children() 재귀만으로도 Toplevel이 화면에 나타날 수 있음
                if child.winfo_class() == 'Toplevel':
                    continue
                self._bind_recursive(child, target_container)
        except:
            pass

    def hide_widget(self, event, widget=None):
        """Hide a widget (Ctrl + Right Click)"""
        if self.layout_locked:
            return "break"
            
        if widget is None:
            widget = event.widget
            
        # [STABILITY FIX] Prevent hiding core widgets
        if hasattr(widget, '_config_key') and widget._config_key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
            messagebox.showwarning("보호됨", f"'{widget._config_key}' 항목은 화면의 필수 구성 요소이므로 숨길 수 없습니다.")
            return "break"
        
        # Remove from place layout
        if widget.winfo_manager() == 'place':
            widget.place_forget()
        elif widget.winfo_manager() == 'grid':
            widget.grid_forget()
            
        # Remove placeholder if exists
        self._remove_placeholder(widget)
        
        # Mark as hidden in config
        if hasattr(widget, '_config_key') and widget._config_key:
            try:
                import json
                if os.path.exists(self.config_path):
                    with open(self.config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                else:
                    config = {}
                
                if 'draggable_geometries' not in config:
                    config['draggable_geometries'] = {}
                
                if widget._config_key not in config['draggable_geometries']:
                    config['draggable_geometries'][widget._config_key] = {}
                
                config['draggable_geometries'][widget._config_key]['hidden'] = True
                
                with open(self.config_path, 'w', encoding='utf-8') as f:
                    json.dump(config, f, ensure_ascii=False, indent=2)
            except Exception as e:
                print(f"Failed to save hide status: {e}")
        return "break"

    def make_draggable(self, widget, config_key=None):
        """Make a widget draggable and resizable with right mouse button"""
        widget._config_key = config_key
        # Save original grid info for reset/placeholder
        # We need to do this immediately while it's still in the grid
        widget._original_grid_info = widget.grid_info()
        
        # Recursively bind to ensure clicking anywhere works
        self._bind_recursive(widget, widget)

    def make_header_draggable(self, widget, target_container):
        """Make a specific widget (header/label) draggable with Left Mouse Button targeting a container"""
        from functools import partial
        widget.bind("<Button-1>", partial(self.on_drag_start, widget=target_container))
        widget.bind("<B1-Motion>", partial(self.on_mouse_motion, widget=target_container))
        widget.bind("<ButtonRelease-1>", partial(self.on_drag_stop, widget=target_container))
        
    def reset_widget_position(self, event, widget=None):
        """Reset widget to original grid position"""
        if self.layout_locked:
            return "break"
            
        if widget is None:
            widget = event.widget
        
        # Remove from place layout
        widget.place_forget()
        
        # Remove placeholder if exists
        self._remove_placeholder(widget)
            
        # Restore to grid
        if hasattr(widget, '_original_grid_info'):
            widget.grid(**widget._original_grid_info)
        
        # Reset size variables if any
        if hasattr(widget, '_start_width'): del widget._start_width
        if hasattr(widget, '_start_height'): del widget._start_height
        
        # Remove from config
        if hasattr(widget, '_config_key') and widget._config_key:
            try:
                import json
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                
                if 'draggable_geometries' in config and widget._config_key in config['draggable_geometries']:
                    del config['draggable_geometries'][widget._config_key]
                
                # Backward compatibility: also check top level
                if widget._config_key in config:
                    del config[widget._config_key]
                    
                with open(self.config_path, 'w', encoding='utf-8') as f:
                    json.dump(config, f, ensure_ascii=False, indent=2)
                    
            except Exception as e:
                print(f"Failed to reset config: {e}")

    def reset_all_widgets_layout(self):
        """Reset all widgets to their original grid slots, unhide them, and reset paned window sashes"""
        if messagebox.askyesno("초기화", "모든 항목의 위치와 크기를 초기화하고 숨겨진 항목을 다시 표시하시겠습니까?\n(창 분할 위치도 초기화됩니다.)"):
            # 1. Reset draggable widgets
            for key, widget in list(self.draggable_items.items()):
                # Explicitly unhide if it was hidden
                if widget.winfo_manager() == '':
                    # For grid, we just grid it back
                    if hasattr(widget, '_original_grid_info'):
                         widget.grid(**widget._original_grid_info)
                
                self.reset_widget_position(None, widget=widget)
            
            # 1.05 [NEW] Use custom defaults if they exist
            if os.path.exists(self.config_path):
                try:
                    with open(self.config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                    
                    custom_defaults = config.get('custom_default_geometries', {})
                    if custom_defaults:
                        # Clear existing geometries to ensure fresh start
                        if 'draggable_geometries' not in config: config['draggable_geometries'] = {}
                        
                        for key, geo in custom_defaults.items():
                            if key in self.draggable_items:
                                w = self.draggable_items[key]
                                w.place(x=geo['x'], y=geo['y'], width=geo['width'], height=geo['height'])
                                # Also update config so it persists as the new current layout
                                config['draggable_geometries'][key] = geo
                        
                        # Save the updated config back to disk and memory
                        with open(self.config_path, 'w', encoding='utf-8') as f:
                            json.dump(config, f, ensure_ascii=False, indent=2)
                        self.tab_config = config
                except Exception as e:
                    print(f"Error applying custom defaults: {e}")
                
            # 1.1 Restore parent propagation safely and reset size
            if hasattr(self, 'entry_inner_frame'):
                self.entry_inner_frame.pack_propagate(True)
                self.entry_inner_frame.grid_propagate(True)
                self.entry_inner_frame.config(height=0, width=0) # Let it auto-size for reset
                
                # Perform a layout update pass
                self.root.update_idletasks()
                
                # Re-apply stability locks
                self._adjust_parent_height(self.entry_inner_frame, force=True)
                self.entry_inner_frame.pack_propagate(False)
                self.entry_inner_frame.grid_propagate(False)
                
                # Refresh scrollregion if in canvas
                if hasattr(self, 'entry_canvas'):
                    self.entry_canvas.configure(scrollregion=self.entry_canvas.bbox("all"))
            
            # 2. Reset sash positions (splitters)
            try:
                if hasattr(self, 'daily_usage_paned'):
                    # Give more space to the entry form by default (500px)
                    self.daily_usage_paned.sashpos(0, 500) 
                if hasattr(self, 'daily_history_paned'):
                    total_w = self.daily_history_paned.winfo_width()
                    if total_w > 100:
                        self.daily_history_paned.sashpos(0, int(total_w * 0.7))
                    else:
                        self.daily_history_paned.sashpos(0, 800)
            except:
                pass

            self.save_tab_config()
            messagebox.showinfo("완료", "레이아웃과 창 분할 위치가 초기화되었습니다.")

    def save_current_layout_as_default(self):
        """Save the current layout as the custom 'reset' default"""
        try:
            current_geos = {}
            for key, widget in self.draggable_items.items():
                # [FIX] Capture all managed widgets (grid or place) 
                # This ensures the snapshot is complete even if some items weren't moved.
                manager = widget.winfo_manager()
                if manager in ['place', 'grid']:
                    current_geos[key] = {
                        'x': widget.winfo_x(),
                        'y': widget.winfo_y(),
                        'width': widget.winfo_width(),
                        'height': widget.winfo_height(),
                        'hidden': False
                    }
            
            if not current_geos:
                messagebox.showwarning("실패", "수동으로 배치된(드래그된) 항목이 없어 기본값으로 저장할 수 없습니다.")
                return

            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            else:
                config = {}
                
            config['custom_default_geometries'] = current_geos
            
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
                
            messagebox.showinfo("완료", "현재 레이아웃이 '기본 배치'로 저장되었습니다.\n앞으로 [초기화] 시 이 배치로 되돌아갑니다.")
        except Exception as e:
            messagebox.showerror("오류", f"기본 배치 저장 중 오류 발생: {e}")
            
    def toggle_layout_lock(self):
        """Toggle layout locking state"""
        self.layout_locked = not self.layout_locked
        if hasattr(self, 'btn_lock_layout'):
            if self.layout_locked:
                self.btn_lock_layout.config(text="🔒 배치 고정됨")
                self.style.configure("Lock.TButton", foreground="black")
            else:
                self.btn_lock_layout.config(text="🔓 배치 수정 중")
                self.style.configure("Lock.TButton", foreground="red")
        
        # Save layout lock state to config immediately
        if not hasattr(self, 'tab_config'):
            self.tab_config = {}
        self.tab_config['layout_locked'] = self.layout_locked
        
        # [FIX] Unlocking is no longer destructive. 
        # It simply enables movement without resetting positions.
        if self.layout_locked:
            print("Layout locked - current positions preserved")
        else:
            print("Layout unlocked - movement enabled")
        
        # Force save the state immediately to file
        self.save_tab_config(force=True)
        
        print(f"Layout lock {'enabled' if self.layout_locked else 'disabled'}")

    def on_drag_stop(self, event, widget=None):
        """Handle end of dragging or resizing and auto-save"""
        if widget is None:
            widget = event.widget
        if hasattr(widget, '_interaction_mode'):
            mode = getattr(widget, '_interaction_mode')
            del widget._interaction_mode
            
            # Update parent height if something moved or resized
            self._adjust_parent_height(widget.master, force=True)
            
            # Auto-save layout
            self.save_tab_config()


    def add_new_memo(self, initial_text="", initial_title="메모", key=None):
        """Create a new movable/editable/copyable memo box with editable title"""
        import time
        if key is None:
            key = f"memo_{int(time.time() * 1000)}"
        
        # Create container
        memo_container = ttk.LabelFrame(self.entry_inner_frame)
        
        # Controls frame (top of memo)
        ctrl_frame = ttk.Frame(memo_container)
        ctrl_frame.pack(fill='x', side='top')
        
        # Editable Title Entry
        title_entry = ttk.Entry(ctrl_frame, font=('Arial', 9, 'bold'), width=15)
        title_entry.pack(side='left', padx=2)
        title_entry.insert(0, initial_title)
        title_entry.bind('<FocusOut>', lambda e: self.save_tab_config())
        
        btn_copy = ttk.Button(ctrl_frame, text="📋", width=3, command=lambda: self.duplicate_memo(key))
        btn_copy.pack(side='right')
        
        btn_del = ttk.Button(ctrl_frame, text="❌", width=3, command=lambda: self.remove_box(key))
        btn_del.pack(side='right')
        
        # Text area
        text_area = tk.Text(memo_container, wrap='word', height=5, width=30, font=('Arial', 10), bg=self.theme_bg, highlightthickness=0)
        text_area.pack(fill='both', expand=True, padx=2, pady=2)
        text_area.insert('1.0', initial_text)
        
        # Bind text change to auto-save
        text_area.bind('<FocusOut>', lambda e: self.save_tab_config())
        
        # Make draggable (Right Click)
        self.make_draggable(memo_container, key)
        
        # Make header draggable (Left Click)
        self.make_header_draggable(ctrl_frame, memo_container)
        self.draggable_items[key] = memo_container
        self.memos[key] = {
            'container': memo_container, 
            'text_widget': text_area,
            'title_entry': title_entry
        }
        
        # Initial placement if not loaded from config (will be handled by load_tab_config if exists)
        if key not in getattr(self, '_loading_memos', []):
            memo_container.place(x=50, y=50) # Default start position
        
        return memo_container

    def duplicate_memo(self, key):
        """Duplicate an existing memo with its title and content"""
        if self.layout_locked: return
        if key in self.memos:
            content = self.memos[key]['text_widget'].get('1.0', 'end-1c')
            title = self.memos[key]['title_entry'].get()
            self.add_new_memo(initial_text=content, initial_title=title)
            self.save_tab_config()

    def add_new_checklist(self, initial_data=None, initial_title="체크리스트", key=None):
        """Create a new movable/editable checklist box"""
        import time
        if key is None:
            key = f"checklist_{int(time.time() * 1000)}"
        
        # Create container
        check_container = ttk.LabelFrame(self.entry_inner_frame)
        
        # Controls frame (top of checklist)
        ctrl_frame = ttk.Frame(check_container)
        ctrl_frame.pack(fill='x', side='top')
        
        # Editable Title Entry
        title_entry = ttk.Entry(ctrl_frame, font=('Arial', 9, 'bold'), width=15)
        title_entry.pack(side='left', padx=2)
        title_entry.insert(0, initial_title)
        title_entry.bind('<FocusOut>', lambda e: self.save_tab_config())
        
        btn_copy = ttk.Button(ctrl_frame, text="📋", width=3, command=lambda: self.duplicate_checklist(key))
        btn_copy.pack(side='right')
        
        btn_del = ttk.Button(ctrl_frame, text="❌", width=3, command=lambda: self.remove_box(key))
        btn_del.pack(side='right')
        
        # Add Item Area
        add_frame = ttk.Frame(check_container)
        add_frame.pack(fill='x', padx=2, pady=2)
        
        new_item_var = tk.StringVar()
        entry_new = ttk.Entry(add_frame, textvariable=new_item_var, width=20)
        entry_new.pack(side='left', fill='x', expand=True)
        
        def add_item(event=None):
             text = new_item_var.get().strip()
             if text:
                 self.add_checklist_item(item_frame, text, False, key)
                 new_item_var.set("")
                 self.save_tab_config()
        
        entry_new.bind('<Return>', add_item)
        btn_add = ttk.Button(add_frame, text="➕", width=3, command=add_item)
        btn_add.pack(side='right')

        # Scrollable Frame for Items
        canvas_frame = ttk.Frame(check_container)
        canvas_frame.pack(fill='both', expand=True, padx=2, pady=2)
        
        canvas = tk.Canvas(canvas_frame, height=100, width=200, bg=self.theme_bg, highlightthickness=0) # Match theme
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        item_frame = ttk.Frame(canvas)
        
        item_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=item_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Scroll binding is handled globally in MaterialManager.__init__

        # Make draggable
        self.make_draggable(check_container, key)
        self.draggable_items[key] = check_container
        self.checklists[key] = {
            'container': check_container,
            'title_entry': title_entry,
            'item_frame': item_frame,
            'items': [] # List of item widgets/vars will be managed dynamically via children of item_frame
        }
        
        # Add initial items if provided
        if initial_data:
            for item in initial_data:
                self.add_checklist_item(item_frame, item.get('text', ''), item.get('checked', False), key)
        
        # Initial placement
        if key not in getattr(self, '_loading_memos', []): # Reuse loading flag or logic
            check_container.place(x=50, y=150)
            
        return check_container

    def add_vehicle_inspection_box(self, initial_data=None, key=None):
        """Create a new movable vehicle inspection box"""
        import time
        if key is None:
            key = f"vehicle_inspection_{int(time.time() * 1000)}"
            
        # Create draggable container
        container, widget = self.create_draggable_container(
            self.entry_inner_frame, "차량 점검", VehicleInspectionWidget, key,
            manage_list_key='vehicles',
            theme_bg=getattr(self, 'theme_bg', '#f0f0f0'),
            vehicle_list=getattr(self, 'vehicles', [])
        )
        
        # Track specifically for data gathering
        self.vehicle_inspections[key] = widget
        
        # Load initial data if provided
        if initial_data:
            widget.set_data(initial_data)
            
        # Initial placement if new
        if key not in getattr(self, '_loading_memos', []):
            container.place(x=50, y=300)
            
        return container

    def add_checklist_item(self, parent_frame, text, checked, checklist_key):
        """Add a single item row to the checklist"""
        row_frame = ttk.Frame(parent_frame)
        row_frame.pack(fill='x', pady=1)
        
        var = tk.BooleanVar(value=checked)
        cb = ttk.Checkbutton(row_frame, variable=var, command=lambda: self.save_tab_config())
        cb.pack(side='left')
        
        entry = ttk.Entry(row_frame)
        entry.insert(0, text)
        entry.pack(side='left', fill='x', expand=True)
        entry.bind('<FocusOut>', lambda e: self.save_tab_config())
        
        def delete_this_item():
            row_frame.destroy()
            self.save_tab_config()
            
        btn_del_item = ttk.Label(row_frame, text="❌", font=('Arial', 8), cursor='hand2', foreground='gray')
        btn_del_item.pack(side='right', padx=2)
        btn_del_item.bind('<Button-1>', lambda e: delete_this_item())
        
        # Store refs in widget for retrieval during save
        row_frame._checklist_data = {'var': var, 'entry': entry}

    def duplicate_checklist(self, key):
        """Duplicate an existing checklist"""
        if self.layout_locked: return
        if key in self.checklists:
            # get current data
            original_items = []
            item_frame = self.checklists[key]['item_frame']
            for child in item_frame.winfo_children():
                if hasattr(child, '_checklist_data'):
                    data = child._checklist_data
                    original_items.append({
                        'text': data['entry'].get(),
                        'checked': data['var'].get()
                    })
            
            title = self.checklists[key]['title_entry'].get()
            self.add_new_checklist(initial_data=original_items, initial_title=title)
            self.save_tab_config()


    def on_drag_start(self, event, widget=None):
        """Begin dragging widget"""
        if self.layout_locked:
            return "break" # Prevent movement and stop propagation
            
        if widget is None:
            widget = event.widget
        widget._interaction_mode = 'move'
        
        # Save absolute start position of mouse
        widget._drag_start_root_x = event.x_root
        widget._drag_start_root_y = event.y_root
        
        # Save initial widget position relative to parent
        widget._drag_start_pos_x = widget.winfo_x()
        widget._drag_start_pos_y = widget.winfo_y()
        
        # Ensure we have grid info (redundant but safe)
        if not hasattr(widget, '_original_grid_info') and widget.grid_info():
            widget._original_grid_info = widget.grid_info()
        
    def on_resize_start(self, event, widget=None):
        """Begin resizing widget"""
        if self.layout_locked:
            return "break"
            
        if widget is None:
            widget = event.widget
        widget._interaction_mode = 'resize'
        
        # Save absolute start position of mouse
        widget._drag_start_root_x = event.x_root
        widget._drag_start_root_y = event.y_root
        
        # Save initial size
        widget._start_width = widget.winfo_width()
        widget._start_height = widget.winfo_height()
        
        # Ensure we have grid info
        if not hasattr(widget, '_original_grid_info') and widget.grid_info():
            widget._original_grid_info = widget.grid_info()
        return "break"
        
    def on_mouse_motion(self, event, widget=None):
        """Handle dragging or resizing motion with performance throttling"""
        if self.layout_locked:
            return "break"
            
        if widget is None:
            widget = event.widget
        
        if not hasattr(widget, '_interaction_mode'):
            return
        
        # PERFORMANCE THROTTLE: Limit updates to ~60fps (16ms)
        curr_time = time.time()
        if curr_time - self._last_motion_time < 0.016:
            # Still update the physical position of the widget being interacted with
            # or the user will feel lag in the initial drag/resize itself.
            self._update_widget_position(event, widget)
            return
        
        self._last_motion_time = curr_time
        
        # Apply positioning only (No collision, No auto-resize)
        self._update_widget_position(event, widget)

    def _update_widget_position(self, event, widget):
        """Internal helper to calculate and set widget position/size during interaction"""
        dx = event.x_root - widget._drag_start_root_x
        dy = event.y_root - widget._drag_start_root_y
        parent = widget.master
        parent_w = parent.winfo_width()
        parent_h = parent.winfo_height()
        
        if widget._interaction_mode == 'move':
            new_x = widget._drag_start_pos_x + dx
            new_y = widget._drag_start_pos_y + dy
            widget_w = widget.winfo_width()
            widget_h = widget.winfo_height()
            
            # Clamp
            if new_x < 0: new_x = 0
            elif new_x + widget_w > parent_w: new_x = max(0, parent_w - widget_w)
            
            # Loosen Y clamping: allow moving slightly below parent to trigger growth
            if new_y < 0: new_y = 0
            # Instead of strict clamp at parent_h, allow a small overflow to trigger resize logic
            elif new_y + widget_h > parent_h + 100: 
                new_y = parent_h + 100 - widget_h

            
            if widget.winfo_manager() != 'place':
                # Save original grid info and explicitly un-grid
                if not hasattr(widget, '_original_grid_info') and widget.grid_info():
                    widget._original_grid_info = widget.grid_info()
                self._ensure_placeholder(widget)
                widget.grid_forget() # Explicitly un-grid before placing
                widget.place(width=widget_w, height=widget_h)
                
            widget.place(x=new_x, y=new_y)
            widget.lift()
            
            # Auto-update scroll region if we're dragging near bounds
            self._ensure_canvas_scroll_region()
            
        elif widget._interaction_mode == 'resize':
            new_width = max(50, widget._start_width + dx)
            new_height = max(20, widget._start_height + dy)
            current_x = widget.winfo_x()
            current_y = widget.winfo_y()
            
            # Clamp
            if current_x + new_width > parent_w: new_width = max(50, parent_w - current_x)
            if current_y + new_height > parent_h: new_height = max(20, parent_h - current_y)
            
            if widget.winfo_manager() != 'place':
                # Save original grid info and explicitly un-grid
                if not hasattr(widget, '_original_grid_info') and widget.grid_info():
                    widget._original_grid_info = widget.grid_info()
                self._ensure_placeholder(widget)
                widget.grid_forget() # Explicitly un-grid before placing
                widget.place(x=current_x, y=current_y)
                
            widget.place(width=new_width, height=new_height)
            widget.lift()
        
        return "break"
        
    def _apply_push_down_logic(self, dragged_widget):
        """Recursively push widgets down if they overlap with the dragged widget"""
        try:
            # Current bounding box of the dragging widget
            x1 = dragged_widget.winfo_x()
            y1 = dragged_widget.winfo_y()
            w1 = dragged_widget.winfo_width()
            h1 = dragged_widget.winfo_height()
            
            padding = 10
            
            # Parent of the widgets
            parent = dragged_widget.master
            
            # OPTIMIZATION: Early filter candidates by parent to avoid massive iterations
            candidates = [w for w in self.draggable_items.values() if w.master == parent]
            
            for other in candidates:
                if other == dragged_widget:
                    continue
                
                # [STABILITY FIX] Skip if the 'other' widget's BOTTOM is ABOVE the dragged widget's TOP.
                # Strictly ignore anything physically above the current interaction zone.
                if (other.winfo_y() + other.winfo_height()) <= y1:
                    continue

                # Skip if not in the same parent
                if other.master != parent:
                    continue
                
                # Skip if hidden
                if not other.winfo_ismapped():
                    continue
                
                # Get current pos/size of the 'other' widget
                x2 = other.winfo_x()
                y2 = other.winfo_y()
                w2 = other.winfo_width()
                h2 = other.winfo_height()
                
                # 1. Horizontal Overlap Check
                if (x1 < x2 + w2) and (x1 + w1 > x2):
                    # 2. Vertical Collision Check
                    # If dragged_widget is hitting 'other' from the top
                    # We check if the bottom of dragged_widget is below the top of other
                    if y1 < y2 and (y1 + h1) > y2:
                        # New target Y for 'other'
                        new_y2 = y1 + h1 + padding
                        
                        # Only move if it's actually pushing it DOWN
                        if y2 < new_y2:
                            if other.winfo_manager() == 'grid':
                                # Convert grid to place
                                self._ensure_placeholder(other)
                                other.lift()
                                other.place(x=x2, y=new_y2, width=w2, height=h2)
                            else:
                                # Just update place
                                other.place(y=new_y2)
                            
                            # 3. Recursive Push
                            self._apply_push_down_logic(other)
        except Exception as e:
            # Silent fail to avoid interrupting drag
            print(f"Error in push-down logic: {e}")

    def _adjust_parent_height(self, parent, force=False):
        """Adjust parent frame height with performance check"""
        try:
            # Only update idletasks if forced or we're not in the middle of a high-speed interaction
            # This is the single biggest cause of UI stutter.
            if force:
                parent.update_idletasks()
            
            # 2. Start with the bounding box of all GRIDDED items
            try:
                # grid_bbox returns (x, y, width, height) of the grid
                bbox = parent.grid_bbox()
                required_h = bbox[1] + bbox[3] if bbox[3] > 0 else 0
                required_w = bbox[0] + bbox[2] if bbox[2] > 0 else 0
            except:
                required_h = 0
                required_w = 0

            # 3. Handle PLACE items as well (dragged/custom items)
            for child in parent.winfo_children():
                try:
                    manager = child.winfo_manager()
                    if manager == 'place':
                        info = child.place_info()
                        # Get relative or absolute y + height
                        y = child.winfo_y()
                        h = child.winfo_height()
                        x = child.winfo_x()
                        w = child.winfo_width()
                        required_h = max(required_h, y + h)
                        required_w = max(required_w, x + w)
                except:
                    continue
            
            # Add some padding
            new_height = required_h + 30
            
            # [STABILITY FIX] Guard against collapse but allow growth
            # Width is now dictated by canvas parent in responsive mode.
            if new_height < 500: new_height = 800
            
            # Only resize if the required dimensions are significantly different
            current_h = parent.winfo_height()
            if abs(current_h - new_height) > 10:
                 parent.config(height=new_height)
                 
                 # [RECURSIVE FIX] If this parent has a master that is part of the entry system,
                 # adjust its height too. 
                 if parent.master and parent.master != self.entry_inner_frame.master: # Don't go above canvas
                     self._adjust_parent_height(parent.master, force=force)
                     
        except Exception as e:
            print(f"Error adjusting parent height: {e}")


    def _ensure_placeholder(self, widget, width=None, height=None):
        """Ensure a placeholder exists in the grid where the widget used to be"""
        # Map widget to a placeholder attribute name dynamically
        # We can use the widget id or a dictionary, but simpler to attach to widget
        if not hasattr(widget, '_placeholder'):
            # Use provided dims or current widget dims
            w = width if width is not None else widget.winfo_width()
            h = height if height is not None else widget.winfo_height()
            
            # Create a frame to hold the space
            widget._placeholder = ttk.Frame(widget.master, width=w, height=h)
            
            # Grid it at the original position
            if hasattr(widget, '_original_grid_info'):
                widget._placeholder.grid(**widget._original_grid_info)
                # Ensure the placeholder doesn't shrink and holds its size
                widget._placeholder.grid_propagate(False)
                
    def _remove_placeholder(self, widget):
        """Remove placeholder for a widget"""
        if hasattr(widget, '_placeholder'):
            widget._placeholder.destroy()
            del widget._placeholder

        

    def setup_daily_usage_tab(self):
        """Setup the daily usage entry tab"""
        # Top frame for entry form
        # We use a canvas or large frame to allow free movement? 
        # Actually, we keep the entry_frame as the parent but use grid for initial layout
        # The user can then move them out of grid into place
        
        # Create PanedWindow for resizable frames
        self.daily_usage_paned = ttk.Panedwindow(self.tab_daily_usage, orient='vertical')
        self.daily_usage_paned.pack(fill='both', expand=True, padx=5, pady=5)  # Reduced padding
        
        # Save sash position on adjustment and lock it
        self.daily_usage_paned.bind("<ButtonRelease-1>", self._on_daily_usage_sash_changed)
        self.daily_usage_paned.bind("<Configure>", self._on_daily_usage_resize)
        # [FIX] Respect loaded config if available, otherwise default to False
        self.daily_usage_sash_locked = getattr(self, 'daily_usage_sash_locked', False)
        
        # Set initial sash position to ensure visibility (30% for top frame, 70% for bottom)
        self.daily_usage_paned.after(200, self._ensure_daily_usage_sash_visibility)
        self.daily_usage_paned.after(500, self._ensure_daily_usage_sash_visibility)
        self.daily_usage_paned.after(1000, self._ensure_daily_usage_sash_visibility)
        self.daily_usage_paned.after(1200, self._ensure_canvas_scroll_region)

        
        entry_frame = ttk.LabelFrame(self.daily_usage_paned, text="현장별 일일 사용량 기입")
        self.daily_usage_paned.add(entry_frame, weight=1) # Changed from weight=3 to weight=1
        
        # Header for buttons to keep them separate from draggable area
        header_frame = ttk.Frame(entry_frame)
        header_frame.pack(fill='x', padx=2, pady=1)  # Reduced padding
        
        btn_reset_all = ttk.Button(header_frame, text="전체 레이아웃 초기화", command=self.reset_all_widgets_layout)
        btn_reset_all.pack(side='right', padx=5)
        
        lock_text = "🔒 배치 고정됨" if getattr(self, 'layout_locked', False) else "🔓 배치 수정 중"
        lock_style = "Lock.TButton" if getattr(self, 'layout_locked', False) else "Lock.TButton"
        self.btn_lock_layout = ttk.Button(header_frame, text=lock_text, command=self.toggle_layout_lock, style=lock_style)
        self.btn_lock_layout.pack(side='right', padx=5)
        if getattr(self, 'layout_locked', False):
             self.style.configure("Lock.TButton", foreground="black")
        else:
             self.style.configure("Lock.TButton", foreground="red")
        
        # Add sash lock toggle button to header
        button_text = "🔒 경계 고정됨" if getattr(self, 'daily_usage_sash_locked', False) else "🔓 경계 고정"
        button_style = "SashLock.TButton" if getattr(self, 'daily_usage_sash_locked', False) else "TButton"
        self.btn_sash_lock = ttk.Button(header_frame, text=button_text, command=self.toggle_sash_lock, style=button_style)
        self.btn_sash_lock.pack(side='right', padx=5)
        
        # Start monitoring immediately if locked
        if getattr(self, 'daily_usage_sash_locked', False):
            self.root.after(1000, self._start_sash_monitor)
        
        # Add resolution display
        self.resolution_label = ttk.Label(header_frame, text="", font=('Malgun Gothic', 8))
        self.resolution_label.pack(side='right', padx=10)
        
        # Add resolution lock button
        self.btn_resolution_lock = ttk.Button(header_frame, text="🔓 해상도 고정", command=self.toggle_resolution_lock)
        self.btn_resolution_lock.pack(side='right', padx=5)
        # [FIX] Initialize from config if available (will be updated further in load_tab_config)
        self.resolution_locked = getattr(self, 'resolution_locked', False)
        
        # Update resolution display
        self.update_resolution_display()
        
        
        btn_add_memo = ttk.Button(header_frame, text="➕ 메모 추가", command=self.add_new_memo)
        btn_add_memo.pack(side='right', padx=5)

        btn_add_checklist = ttk.Button(header_frame, text="☑️ 목록 추가", command=self.add_new_checklist)
        btn_add_checklist.pack(side='right', padx=5)
        
        btn_add_vehicle = ttk.Button(header_frame, text="🚗 차량점검", command=self.add_vehicle_inspection_box)
        btn_add_vehicle.pack(side='right', padx=5)
        
        btn_save_default = ttk.Button(header_frame, text="⭐ 기본 배치로 저장", command=self.save_current_layout_as_default)
        btn_save_default.pack(side='right', padx=5)

        # [STABILITY FIX] Use a Canvas to hold the entry form for draggable support.
        # [REVISION] Scrollbars removed per user request for a cleaner UI.
        canvas_parent = ttk.Frame(entry_frame)
        canvas_parent.pack(fill='both', expand=True, padx=2, pady=1)
        
        self.entry_canvas = tk.Canvas(canvas_parent, highlightthickness=0, bg=self.theme_bg)
        # Scrollbars (entry_vsb/hsb) and their configurations have been removed.
        self.entry_canvas.pack(side='left', fill='both', expand=True)
        
        self.entry_inner_frame = ttk.Frame(self.entry_canvas)
        # Use a window to place the frame inside the canvas
        self.entry_canvas_window = self.entry_canvas.create_window((0, 0), window=self.entry_inner_frame, anchor='nw')
        
        def _on_entry_config(e):
            # Update canvas window width: use canvas width but ensure a minimum of 1100 
            # to fit both form (550) and workers (550+) without clipping.
            target_w = max(1100, e.width)
            self.entry_canvas.itemconfig(self.entry_canvas_window, width=target_w)
            self._ensure_canvas_scroll_region()
        
        self.entry_inner_frame.bind("<Configure>", lambda e: self._ensure_canvas_scroll_region())
        self.entry_canvas.bind("<Configure>", _on_entry_config)
        
        # Mouse Wheel bindings have been removed to prevent scrolling as requested.
        
        # Explicitly fix all possible grid rows to weight 0
        for r in range(100):
            self.entry_inner_frame.grid_rowconfigure(r, weight=0)
        
        # Configure the main 2nd column container (Workers side by default)
        # Give column 0 (Forms) a fixed minimum width to prevent crushing
        self.entry_inner_frame.grid_columnconfigure(0, weight=0, minsize=550)
        self.entry_inner_frame.grid_columnconfigure(1, weight=1)
        self.entry_inner_frame.grid_rowconfigure(0, weight=1)

        # 1. Combined Frame: Basic Info & Costs
        form_container = ttk.Frame(self.entry_inner_frame, relief="solid", borderwidth=1)
        form_hdr = ttk.Frame(form_container)
        form_hdr.pack(fill='x', side='top')
        ttk.Label(form_hdr, text="✥", font=('Arial', 9)).pack(side='left', padx=1)
        lbl_form_title = ttk.Label(form_hdr, text="기본 정보 및 검사 비용", font=('Malgun Gothic', 9, 'bold'))
        lbl_form_title.pack(side='left', padx=1)
        
        # Inner content for the form
        form_content = ttk.Frame(form_container, padding=(1, 0))
        form_content.pack(fill='both', expand=True)
        
        # Row 0: Site & Date
        ttk.Label(form_content, text="현장명:").grid(row=0, column=0, padx=2, pady=1, sticky='e')
        self.cb_daily_site = ttk.Combobox(form_content, width=12)
        self.cb_daily_site.grid(row=0, column=1, padx=2, pady=1, sticky='w')
        register_autocomplete(self.cb_daily_site, None) # Use general registration
        
        # [NEW] Focus Transition for Site
        def on_site_select(e):
            self.root.after(10, self.ent_daily_date.focus_set)
            return "break"
        self.cb_daily_site.bind('<<ComboboxSelected>>', on_site_select)
        self.cb_daily_site.bind('<Return>', on_site_select)

        # [NEW] 우클릭 컨텍스트 메뉴: 현장명 목록 관리
        def _show_site_context_menu(e=None):
            site_val = self.cb_daily_site.get().strip()
            menu = tk.Menu(self.root, tearoff=0)
            if site_val:
                if site_val in getattr(self, 'hidden_sites', []):
                    menu.add_command(
                        label=f'✅ "{site_val}" 목록에 다시 표시',
                        command=lambda: self._toggle_hidden_site(site_val, hide=False)
                    )
                else:
                    menu.add_command(
                        label=f'🚫 "{site_val}" 목록에서 숨기기',
                        command=lambda: self._toggle_hidden_site(site_val, hide=True)
                    )
                menu.add_separator()
            menu.add_command(label='숨긴 현장 모두 표시', command=self._show_all_hidden_sites)
            # 버튼 위치 기준으로 팝업
            try:
                x = btn_site_mgr.winfo_rootx()
                y = btn_site_mgr.winfo_rooty() + btn_site_mgr.winfo_height()
                menu.tk_popup(x, y)
            finally:
                menu.grab_release()

        btn_site_mgr = ttk.Button(form_content, text="⚙", width=2,
                                   command=_show_site_context_menu)
        btn_site_mgr.grid(row=0, column=1, padx=(95, 0), pady=1, sticky='w')
        # 우클릭도 동일하게 동작
        for w in [self.cb_daily_site, btn_site_mgr]:
            w.bind('<Button-3>', _show_site_context_menu)


        ttk.Label(form_content, text="날짜:").grid(row=0, column=2, padx=2, pady=1, sticky='e')
        self.ent_daily_date = DateEntry(form_content, width=12, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        self.ent_daily_date.grid(row=0, column=3, padx=2, pady=1, sticky='w')
        
        # [FIX] Focus Transition for Date. Now simplified thanks to the global DateEntry patch.
        def on_date_select_combined(e=None):
            # Focus transition to Equipment field
            self.root.after(100, self.cb_daily_equip.focus_set)
            return "break"
            
        self.ent_daily_date.bind('<<DateEntrySelected>>', on_date_select_combined)
        self.ent_daily_date.bind('<Return>', on_date_select_combined)

        # Row 1: Equipment & Material
        ttk.Label(form_content, text="장비명:").grid(row=1, column=0, padx=2, pady=1, sticky='e')
        self.cb_daily_equip = ttk.Combobox(form_content, width=12)
        self.cb_daily_equip.grid(row=1, column=1, padx=2, pady=1, sticky='w')
        register_autocomplete(self.cb_daily_equip, None)
        
        # [NEW] Focus Transition for Equipment
        def on_equip_select(e):
            self.root.after(10, self.cb_daily_material.focus_set)
            return "break"
        self.cb_daily_equip.bind('<<ComboboxSelected>>', on_equip_select)
        self.cb_daily_equip.bind('<Return>', on_equip_select)
        
        ttk.Label(form_content, text="품목명:").grid(row=1, column=2, padx=2, pady=1, sticky='e')
        self.cb_daily_material = ttk.Combobox(form_content, width=12)
        self.cb_daily_material.grid(row=1, column=3, padx=2, pady=1, sticky='w')
        register_autocomplete(self.cb_daily_material, None)
        
        # [NEW] Focus Transition for Material
        def on_mat_select(e):
            self.root.after(10, self.cb_daily_test_method.focus_set)
            return "break"
        self.cb_daily_material.bind('<<ComboboxSelected>>', on_mat_select)
        self.cb_daily_material.bind('<Return>', on_mat_select)
        
        # Row 2: Method & Quantity
        ttk.Label(form_content, text="방법:").grid(row=2, column=0, padx=2, pady=1, sticky='e')
        self.cb_daily_test_method = ttk.Combobox(form_content, width=12, values=["RT", "PAUT", "UT", "MT", "PT", "PMI"])
        self.cb_daily_test_method.grid(row=2, column=1, padx=2, pady=1, sticky='w')
        
        # [NEW] Focus Transition for Method
        def on_method_select(e):
            self.root.after(10, self.ent_daily_test_amount.focus_set)
            return "break"
        self.cb_daily_test_method.bind('<<ComboboxSelected>>', on_method_select, add='+')
        self.cb_daily_test_method.bind('<Return>', on_method_select, add='+')
        
        ttk.Label(form_content, text="수량:").grid(row=2, column=2, padx=2, pady=1, sticky='e')
        self.ent_daily_test_amount = ttk.Entry(form_content, width=12)
        self.ent_daily_test_amount.grid(row=2, column=3, padx=2, pady=1, sticky='w')
        
        # Row 3: Costs
        ttk.Label(form_content, text="단가:").grid(row=3, column=0, padx=2, pady=1, sticky='e')
        self.ent_daily_unit_price = ttk.Entry(form_content, width=12)
        self.ent_daily_unit_price.grid(row=3, column=1, padx=2, pady=1, sticky='w')
        
        ttk.Label(form_content, text="출장비:").grid(row=3, column=2, padx=2, pady=1, sticky='e')
        self.ent_daily_travel_cost = ttk.Entry(form_content, width=12)
        self.ent_daily_travel_cost.grid(row=3, column=3, padx=2, pady=1, sticky='w')
        
        ttk.Label(form_content, text="식대:").grid(row=4, column=0, padx=2, pady=1, sticky='e')
        self.ent_daily_meal_cost = ttk.Entry(form_content, width=12)
        self.ent_daily_meal_cost.grid(row=4, column=1, padx=2, pady=1, sticky='w')
        
        ttk.Label(form_content, text="검사비:").grid(row=4, column=2, padx=2, pady=1, sticky='e')
        self.ent_daily_test_fee = ttk.Entry(form_content, width=12)
        self.ent_daily_test_fee.grid(row=4, column=3, padx=2, pady=1, sticky='w')
        
        ttk.Label(form_content, text="필름:").grid(row=5, column=0, padx=2, pady=1, sticky='e')
        self.ent_film_count = ttk.Entry(form_content, width=12)
        self.ent_film_count.grid(row=5, column=1, padx=2, pady=1, sticky='w')
        
        # Add Sync button at the end of cost row for better flow
        btn_sync = ttk.Button(form_content, text="일괄", command=self.sync_worker_times, width=4)
        btn_sync.grid(row=5, column=3, padx=(10, 1), sticky='w')
        
        form_container.grid(row=0, column=0, padx=1, pady=1, sticky='nw')
        self.make_draggable(form_container, 'form_box_geometry')
        self.draggable_items['form_box_geometry'] = form_container
        
        # Category definitions for focus flow and loops
        ndt_materials = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
        rtk_cats = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타", "총계"]

        # Row 1: NDT
        ndt_container = ttk.Frame(self.entry_inner_frame, relief="solid", borderwidth=1)
        ndt_hdr = ttk.Frame(ndt_container)
        ndt_hdr.pack(fill='x', side='top'); ttk.Label(ndt_hdr, text="✥", font=('Arial', 9)).pack(side='left', padx=1)
        lbl_ndt_title = ttk.Label(ndt_hdr, text="NDT 자재", font=('Malgun Gothic', 9, 'bold'))
        lbl_ndt_title.pack(side='left', padx=1); ndt_content = ttk.Frame(ndt_container, padding=(1, 0))
        ndt_content.pack(fill='both', expand=True); ndt_grid = ttk.LabelFrame(ndt_content)
        ndt_grid.pack(fill='both', expand=True, padx=1, pady=1)
        
        self.ndt_entries = {}
        for i, mat in enumerate(ndt_materials):
            r = i // 4; col = (i % 4) * 2
            ttk.Label(ndt_grid, text=f"{mat}:", font=('Arial', 8)).grid(row=r, column=col, padx=1, pady=1, sticky='w')
            e = ttk.Entry(ndt_grid, width=6); e.grid(row=r, column=col+1, padx=1, pady=1, sticky='ew'); self.ndt_entries[mat] = e
            
            # Focus transition to next NDT entry
            if i + 1 < len(ndt_materials):
                next_mat = ndt_materials[i+1]
                e.bind('<Return>', lambda e, nm=next_mat: self.ndt_entries[nm].focus_set())
            elif len(rtk_cats) > 0: # From last NDT to first RTK
                e.bind('<Return>', lambda e: self.rtk_entries[rtk_cats[0]].focus_set())

        ndt_container.grid(row=1, column=0, padx=1, pady=1, sticky='nw')
        ndt_container.lift() # Ensure visibility
        self.make_draggable(ndt_container, 'ndt_usage_box_geometry'); self.draggable_items['ndt_usage_box_geometry'] = ndt_container

        # Row 2: RTK
        rtk_container = ttk.Frame(self.entry_inner_frame, relief="solid", borderwidth=1)
        rtk_hdr = ttk.Frame(rtk_container)
        rtk_hdr.pack(fill='x', side='top'); ttk.Label(rtk_hdr, text="✥", font=('Arial', 9)).pack(side='left', padx=1)
        lbl_rtk_title = ttk.Label(rtk_hdr, text="RTK", font=('Malgun Gothic', 9, 'bold'))
        lbl_rtk_title.pack(side='left', padx=1); rtk_content = ttk.Frame(rtk_container, padding=(1, 0))
        rtk_content.pack(fill='both', expand=True); rtk_grid = ttk.LabelFrame(rtk_content)
        rtk_grid.pack(fill='both', expand=True, padx=1, pady=1)
        
        self.rtk_entries = {}
        for i, cat in enumerate(rtk_cats):
            r = i // 4; col = (i % 4) * 2
            ttk.Label(rtk_grid, text=f"{cat}:", font=('Arial', 8)).grid(row=r, column=col, padx=1, pady=1, sticky='w')
            e = ttk.Entry(rtk_grid, width=6)
            e.grid(row=r, column=col+1, padx=1, pady=1, sticky='ew')
            self.rtk_entries[cat] = e
            
            # Focus transition to next RTK entry
            if i + 1 < len(rtk_cats) - 1: # Skip "총계" (the last one)
                next_cat = rtk_cats[i+1]
                e.bind('<Return>', lambda e, nc=next_cat: self.rtk_entries[nc].focus_set())
            elif cat != "총계": # From last editable RTK to save button container (approx)
                # We don't have a direct handle to the button, but we can focus the first worker's name
                e.bind('<Return>', lambda e: self.cb_daily_user.focus_set())
            
            # Bind auto-calculation
            if cat != "총계":
                e.bind('<KeyRelease>', lambda e: self.calculate_rtk_total())
        
        self.rtk_entries["총계"].config(state='readonly')
        rtk_container.grid(row=2, column=0, padx=1, pady=1, sticky='nw')
        rtk_container.lift() # Ensure visibility
        self.make_draggable(rtk_container, 'rtk_usage_box_geometry'); self.draggable_items['rtk_usage_box_geometry'] = rtk_container

        # Row 3: Save Button
        save_btn_container = ttk.Frame(self.entry_inner_frame)
        ttk.Button(save_btn_container, text="주진철 저장", command=self.add_daily_usage_entry, width=15, style='Big.TButton').pack()
        save_btn_container.grid(row=3, column=0, padx=1, pady=1, sticky='nw')
        self.make_draggable(save_btn_container, 'save_btn_geometry'); self.draggable_items['save_btn_geometry'] = save_btn_container

        # [MIGRATION] Convert existing times to "익일" format and sort
        self.worktimes = self._migrate_worktimes(self.worktimes if hasattr(self, 'worktimes') else [])
        
        # Create a single container for all workers
        workers_box, workers_inner = self.create_draggable_container(
            self.entry_inner_frame, "작업자 기록", ttk.Frame, 'workers_box_geometry',
            manage_list_key='users',
            grid_info={'row': 0, 'column': 1, 'rowspan': 4, 'sticky': 'nw', 'padx': 1, 'pady': 1}
        )
        # Configure inner grid for workers (2 columns)
        for c in range(2): workers_inner.grid_columnconfigure(c, weight=1)
        
        def setup_worker_group(idx, row, col):
            # Each worker group is no longer individually draggable
            # Instead, they are sub-frames within the main workers_inner
            group_frame = ttk.LabelFrame(workers_inner, text=f"작업자 {idx}")
            group_frame.grid(row=row, column=col, padx=2, pady=1, sticky='ew')
            
            group = WorkerDataGroup(
                group_frame, worker_index=idx, users_list=getattr(self, 'users', []),
                enable_autocomplete=True,
                time_list=self.worktimes
            )
            group.pack(fill='x', expand=True, padx=1, pady=1)

            # Store global references for all workers (1-10)
            if idx == 1:
                self.cb_daily_user = group.composite
                self.ent_worktime1 = group.ent_worktime
                self.ent_ot1 = group.ent_ot
            
            # Always set index-specific attributes for sync/data access
            setattr(self, f'worker_group{idx}', group)
            setattr(self, f'cb_daily_user{idx}', group.composite)
            setattr(self, f'ent_worktime{idx}', group.ent_worktime)
            setattr(self, f'ent_ot{idx}', group.ent_ot)
            
            # Bindings for auto-save & auto-OT
            group.bind_name('<FocusOut>', lambda e: self.auto_save_to_list(e, group.cb_name, self.users, 'users'))
            group.bind_name('<Return>', lambda e: self.auto_save_to_list(e, group.cb_name, self.users, 'users'))
            
            group.bind_time('<FocusOut>', lambda e: self.auto_save_worktime(e, group.ent_worktime, 'worktimes'))
            group.bind_time('<Return>', lambda e: self.auto_save_worktime(e, group.ent_worktime, 'worktimes'))
            
            group.bind_ot('<FocusOut>', lambda e: self.auto_save_ot(e, group.ent_ot, 'ot_times'))
            group.bind_ot('<Return>', lambda e: self.auto_save_ot(e, group.ent_ot, 'ot_times'))
            
            return group_frame, group

        # 5 rows x 2 cols = 10 workers rearranged inside workers_inner
        for i in range(1, 6): setup_worker_group(i, i-1, 0)
        for i in range(6, 11): setup_worker_group(i, i-6, 1)

        # [FIX] Re-bind drag events recursively so clicking anywhere inside the (now populated) 
        # frames also triggers movement of the main container.
        self.make_draggable(workers_box, 'workers_box_geometry')

        # Bindings & Finalization
        calc_trigger = lambda e: self.update_daily_test_fee_calc()
        
        def on_qty_change(e):
            self.update_daily_test_fee_calc()
            self.sync_film_with_quantity()
            
        # Handle Date Changes globally for recalculations
        def on_date_change(e=None):
            # Recalculate OT for all workers when date changes (weekend vs weekday rates)
            for i in range(1, 11):
                group = getattr(self, f'worker_group{i}', None)
                if group:
                    self.calculate_and_update_ot(group.ent_worktime.get(), group.ent_worktime)

        self.ent_daily_date.bind('<<DateEntrySelected>>', on_date_change, add='+')
        
        # Initial call to set last known date for change detection
        try: self._last_daily_date = self.ent_daily_date.get_date()
        except: self._last_daily_date = None
        
        self.ent_daily_test_amount.bind('<KeyRelease>', on_qty_change)
        self.cb_daily_test_method.bind('<<ComboboxSelected>>', self.sync_film_with_quantity, add='+')
        
        # [NEW] Add comma auto-formatting and Focus Transition on FocusOut/Return
        cost_entries = [
            self.ent_daily_test_amount,
            self.ent_daily_unit_price, 
            self.ent_daily_travel_cost, 
            self.ent_daily_meal_cost, 
            self.ent_daily_test_fee
        ]
        
        for i, ent in enumerate(cost_entries):
            ent.bind('<KeyRelease>', calc_trigger, add='+')
            ent.bind('<FocusOut>', lambda e, widget=ent: self.format_entry_with_commas(e, widget), add='+')
            
            # Define focus transition
            def on_return(e, current_idx=i):
                # Format current entry first
                self.format_entry_with_commas(e, cost_entries[current_idx])
                # Move focus to next entry if possible
                if current_idx + 1 < len(cost_entries):
                    cost_entries[current_idx + 1].focus_set()
                return "break"
                
            ent.bind('<Return>', on_return)
        
        self.update_material_combo()
        self._adjust_parent_height(self.entry_inner_frame, force=True)
        
        display_frame = ttk.LabelFrame(self.daily_usage_paned, text="일일 사용량 기록 조회")
        self.daily_usage_paned.add(display_frame, weight=1) # Less weight for the list
        
        # Filter controls
        filter_frame = ttk.Frame(display_frame)
        filter_frame.pack(fill='x', padx=5, pady=5)
        
        # Create a sub-frame for filters that need autocomplete
        filter_panel = ttk.Frame(filter_frame)
        filter_panel.pack(fill='x', padx=0, pady=0) # Pack below date filters
        
        ttk.Label(filter_frame, text="시작일:").pack(side='left', padx=5)
        self.ent_daily_start_date = DateEntry(filter_frame, width=12, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        self.ent_daily_start_date.pack(side='left', padx=5)
        # Default to 30 days ago
        start_date = (datetime.datetime.now() - datetime.timedelta(days=30))
        self.ent_daily_start_date.set_date(start_date)
        
        ttk.Label(filter_frame, text="종료일:").pack(side='left', padx=5)
        self.ent_daily_end_date = DateEntry(filter_frame, width=12, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        self.ent_daily_end_date.pack(side='left', padx=5)
        self.ent_daily_end_date.set_date(datetime.datetime.now())
        
        ttk.Label(filter_panel, text="현장:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_site = ttk.Combobox(filter_panel, width=12)
        self.cb_daily_filter_site.pack(side='left', padx=2)
        self.cb_daily_filter_site.set('전체')
        
        ttk.Label(filter_panel, text="품목명:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_material = ttk.Combobox(filter_panel, width=15)
        self.cb_daily_filter_material.pack(side='left', padx=2)
        self.cb_daily_filter_material.set('전체')
        
        ttk.Label(filter_panel, text="장비명:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_equipment = ttk.Combobox(filter_panel, width=15)
        self.cb_daily_filter_equipment.pack(side='left', padx=2)
        self.cb_daily_filter_equipment.set('전체')
        
        ttk.Label(filter_panel, text="작업자:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_worker = ttk.Combobox(filter_panel, width=12)
        self.cb_daily_filter_worker.pack(side='left', padx=2)
        self.cb_daily_filter_worker.set('전체')
        
        ttk.Label(filter_panel, text="차량번호:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_vehicle = ttk.Combobox(filter_panel, width=12)
        self.cb_daily_filter_vehicle.pack(side='left', padx=2)
        self.cb_daily_filter_vehicle.set('전체')
        
        ttk.Label(filter_panel, text="분류:").pack(side='left', padx=(5, 2))
        self.cb_daily_filter_shift = ttk.Combobox(filter_panel, width=10, state="readonly", values=["전체", "주간", "야간", "주야간", "휴일"])
        self.cb_daily_filter_shift.pack(side='left', padx=2)
        self.cb_daily_filter_shift.set('전체')
        
        btn_filter = ttk.Button(filter_frame, text="조회", command=self.update_daily_usage_view)
        btn_filter.pack(side='left', padx=10)
        
        btn_delete = ttk.Button(filter_frame, text="선택 항목 삭제", command=self.delete_daily_usage_entry)
        btn_delete.pack(side='left', padx=10)
        
        btn_edit = ttk.Button(filter_frame, text="선택 항목 수정", command=self.open_edit_daily_usage_dialog)
        btn_edit.pack(side='left', padx=5)
        
        btn_export = ttk.Button(filter_frame, text="엑셀 내보내기", command=self.export_daily_usage_history)
        btn_export.pack(side='left', padx=5)
        
        btn_export_all = ttk.Button(filter_frame, text="전체 기록 내보내기", command=self.export_all_daily_usage)
        btn_export_all.pack(side='left', padx=5)
        
        btn_col_manage = ttk.Button(filter_frame, text="컬럼 관리", command=self.show_column_visibility_dialog)
        btn_col_manage.pack(side='left', padx=10)

        # [NEW] Bind Enter key and selection events to all filter widgets for convenience
        filter_widgets = [
            self.cb_daily_filter_site, self.cb_daily_filter_material, 
            self.cb_daily_filter_equipment, self.cb_daily_filter_worker, 
            self.cb_daily_filter_vehicle, self.cb_daily_filter_shift
        ]
        for widget in filter_widgets:
            widget.bind("<Return>", lambda e: self.update_daily_usage_view())
            widget.bind("<<ComboboxSelected>>", lambda e: self.update_daily_usage_view())

        # DateEntry widgets require binding to their internal entry or using their own events
        for date_widget in [self.ent_daily_start_date, self.ent_daily_end_date]:
            # Some versions of tkcalendar.DateEntry allow direct binding, 
            # but binding to the underlying entry is most reliable for <Return>
            try:
                date_widget.bind("<Return>", lambda e: self.update_daily_usage_view())
                # If DateEntry has an internal entry child, bind to it too
                for child in date_widget.winfo_children():
                    if isinstance(child, (tk.Entry, ttk.Entry)):
                        child.bind("<Return>", lambda e: self.update_daily_usage_view())
            except: pass
        
        # Treeview for daily usage records
        tree_container = ttk.Frame(display_frame)
        tree_container.pack(expand=True, fill='both', padx=5, pady=5)
        
        list_frame = ttk.Frame(tree_container)
        list_frame.pack(fill='both', expand=True)

        # Scrollbars
        vsb = ttk.Scrollbar(list_frame, orient="vertical")
        hsb = ttk.Scrollbar(list_frame, orient="horizontal")
        
        # Treeview with RTK categories and NDT materials
        # Note: Workers 1-10 columns are kept in the 'columns' tuple for data storage,
        # but we will only show a consolidated '작업자' in 'displaycolumns'.
        # Added '(Full작업자)' for Excel export backup.
        columns = ('날짜', '현장', '작업자', '작업시간', 'OT1', 'OT2', 'OT3', 'OT4', 'OT5', 'OT6', 'OT7', 'OT8', 'OT9', 'OT10', '장비명', '검사방법', '수량', '필름매수', '단가', '출장비', '일식', '검사비', 'OT시간', 'OT금액', '품목명', '센터미스', '농도', '마킹미스', '필름마크', '취급부주의', '고객불만', '기타', 'RTK총계', '형광자분', '흑색자분', '백색페인트', '침투제', '세척제', '현상제', '형광침투제', '비고', '입력시간', '차량번호', '주행거리', '차량점검', '차량비고', '(Full작업자)')
        self.daily_usage_tree = ttk.Treeview(list_frame, columns=columns, show='headings',
                                              yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Define display columns - hidden (Full작업자) and individual OT columns by default
        # [NEW] Also hide 'OT시간' by default as per amount-only display request
        visible_defaults = ['날짜', '현장', '작업자', '작업시간', '장비명', '검사방법', '수량', '필름매수', '단가', '출장비', '일식', '검사비', 'OT금액', '품목명', 'RTK총계', '차량번호', '주행거리', '차량점검', '차량비고', '입력시간']
        self.daily_usage_tree['displaycolumns'] = visible_defaults
        
        vsb.config(command=self.daily_usage_tree.yview)
        hsb.config(command=self.daily_usage_tree.xview)
        
        # Column configuration
        col_widths = {
            '날짜': 160, '현장': 130, '작업자': 170, '작업시간': 120,
            'OT1': 140, 'OT2': 140, 'OT3': 140, 'OT4': 140, 'OT5': 140, 'OT6': 140,
            'OT7': 140, 'OT8': 140, 'OT9': 140, 'OT10': 140,
            '차량번호': 120, '주행거리': 100, '차량점검': 200, '차량비고': 150,
            '장비명': 130, '검사방법': 90, 
            '수량': 80, '필름매수': 80, '단가': 90, '출장비': 90, '일식': 80, 
            '검사비': 100, 'OT시간': 80, 'OT금액': 100, '품목명': 210, '센터미스': 70, '농도': 70, '마킹미스': 70, 
            '필름마크': 70, '취급부주의': 70, '고객불만': 70, '기타': 70, 'RTK총계': 80, 
            '형광자분': 80, '흑색자분': 80, '백색페인트': 80, '침투제': 80, '세척제': 80, 
            '현상제': 80, '형광침투제': 80, '비고': 230, '입력시간': 300
        }
        
        for col in columns:
            self.daily_usage_tree.heading(col, text=col, command=lambda c=col: self.treeview_sort_column(self.daily_usage_tree, c, False))
            width = col_widths.get(col, 100)
            self.daily_usage_tree.column(col, width=width, minwidth=20, stretch=False, anchor='center')
        
        # Grid layout for list_frame
        self.daily_usage_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)
        
        # Auto-save column widths when user resizes columns
        # [NEW] Bind worker details popup to both click and double-click
        self.daily_usage_tree.bind("<Button-1>", lambda e: self.show_worker_popup(e, self.daily_usage_tree), add="+")
        self.daily_usage_tree.bind("<Double-1>", lambda e: self.show_worker_popup(e, self.daily_usage_tree), add="+")
        
        # [NEW] Enable column reordering via drag & drop
        self.enable_tree_column_drag(self.daily_usage_tree)
        
        def save_column_widths(event=None):
            self.save_tab_config()
            
        self.daily_usage_tree.bind('<ButtonRelease-1>', save_column_widths)

        # Selection binding
        self.daily_usage_tree.bind('<<TreeviewSelect>>', lambda e: None)
        
        # Optimize treeview scroll region to prevent unnecessary scrollbars
        def optimize_treeview_scroll_region():
            try:
                if hasattr(self, 'daily_usage_tree'):
                    self.daily_usage_tree.update_idletasks()
                    
                    # Get treeview content dimensions
                    total_items = len(self.daily_usage_tree.get_children())
                    if total_items > 0:
                        # Calculate approximate content height
                        item_height = 25  # Approximate height per row
                        content_height = total_items * item_height + 50  # Add some padding
                        
                        # Get treeview dimensions
                        tree_width = self.daily_usage_tree.winfo_width()
                        tree_height = self.daily_usage_tree.winfo_height()
                        
                        # Set scroll region to content size if smaller than treeview
                        if content_height < tree_height:
                            self.daily_usage_tree.configure(yscrollcommand=(0, 0, tree_width, content_height))
                            print(f"Treeview scroll region optimized: (0, 0, {tree_width}, {content_height})")
                        else:
                            # Use full treeview dimensions
                            self.daily_usage_tree.configure(yscrollcommand=(0, 0, tree_width, tree_height))
                            print(f"Treeview scroll region set to full size: (0, 0, {tree_width}, {tree_height})")
            except Exception as e:
                print(f"Error optimizing treeview scroll region: {e}")
        
        # Apply optimization after a short delay
        self.root.after(500, optimize_treeview_scroll_region)
        
        # Initial view update - Moved to end of create_widgets

    


    def export_budget_sales_status(self):
        """Export the currently filtered budget sales status to Excel"""
        if not hasattr(self, 'budget_view_tree') or not self.budget_view_tree.get_children():
            messagebox.showwarning("데이터 없음", "내보낼 데이터가 없습니다. 먼저 조회를 해주세요.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"매출현황_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not file_path:
            return

        # Collect data from treeview
        data = []
        columns = [self.budget_view_tree.heading(col)['text'] for col in self.budget_view_tree['columns']]
        for item in self.budget_view_tree.get_children():
            data.append(self.budget_view_tree.item(item)['values'])
            
        df = pd.DataFrame(data, columns=columns)
        
        # Add summary row at the bottom
        summary_row = {col: '' for col in columns}
        summary_row['날짜'] = '합계'
        if hasattr(self, 'lbl_bv_actual_income'):
            summary_row['검사단가'] = self.lbl_bv_actual_income.cget('text').replace(' 원', '').replace(',', '').strip()
        
        # Compute OT/Total directly from columns since Budget KPIs might show differently
        ot_idx = columns.index('OT합계') if 'OT합계' in columns else -1
        total_idx = columns.index('합계') if '합계' in columns else -1
        
        sum_ot = 0
        sum_tot = 0
        for _d in data:
            if ot_idx >= 0:
                try: sum_ot += float(str(_d[ot_idx]).replace(',', ''))
                except: pass
            if total_idx >= 0:
                try: sum_tot += float(str(_d[total_idx]).replace(',', ''))
                except: pass

        summary_row['OT합계'] = f"{sum_ot:g}"
        summary_row['합계'] = f"{sum_tot:g}"
        
        # Add the summary row to the data list for the DataFrame
        data.append([summary_row.get(col, '') for col in columns])
        
        df = pd.DataFrame(data, columns=columns)
        
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='SalesStatus', index=False)
                # Auto-adjust columns
                worksheet = writer.sheets['SalesStatus']
                for i, col in enumerate(df.columns):
                    max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                    worksheet.column_dimensions[chr(65 + i)].width = max_len
            
            messagebox.showinfo("내보내기 완료", "매출현황이 엑셀로 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 저장 중 오류가 발생했습니다: {e}")

    def setup_budget_tab(self):
        """Setup the project execution budget management tab with detailed labor cost breakdown"""
        # KPI Summary Panel [NEW]
        kpi_frame = tk.Frame(self.tab_budget, background="#ffffff", highlightthickness=1, highlightbackground="#e5e7eb")
        kpi_frame.pack(fill='x', padx=10, pady=(5, 0))
        
        kpis = [
            ("도급액", "lbl_kpi_rev", "#374151"),
            ("실행원가", "lbl_kpi_cost", "#374151"),
            ("영업이익", "lbl_kpi_profit", "#10b981"),
            ("이익률", "lbl_kpi_margin", "#10b981")
        ]
        
        for i, (label, attr, color) in enumerate(kpis):
            f = tk.Frame(kpi_frame, background="#ffffff")
            f.pack(side='left', expand=True, fill='both', pady=10)
            if i > 0:
                tk.Frame(kpi_frame, width=1, background="#e5e7eb").pack(side='left', fill='y', pady=10)
            
            tk.Label(f, text=label, font=("Malgun Gothic", 9), background="#ffffff", foreground="#6b7280").pack()
            lbl_val = tk.Label(f, text="0원", font=("Malgun Gothic", 12, "bold"), background="#ffffff", foreground=color)
            lbl_val.pack()
            setattr(self, attr, lbl_val)

        # Fixed Button Frame [NEW]
        top_btn_frame = ttk.Frame(self.tab_budget)
        top_btn_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Button(top_btn_frame, text="실적 데이터로 채우기", command=self.fill_budget_from_actuals).pack(side='left', padx=5)
        ttk.Button(top_btn_frame, text="저장/수정", command=self.save_budget_entry).pack(side='left', padx=5)
        ttk.Button(top_btn_frame, text="삭제", command=self.delete_budget_entry).pack(side='left', padx=5)
        ttk.Button(top_btn_frame, text="초기화", command=self.clear_budget_form).pack(side='left', padx=5)
        ttk.Button(top_btn_frame, text="엑셀 내보내기", command=self.export_budget_sales_status).pack(side='right', padx=10)

        main_paned = ttk.PanedWindow(self.tab_budget, orient='vertical')
        main_paned.pack(fill='both', expand=True, padx=10, pady=(0, 5))
        
        # --- 상단: Scrollable Form area ---
        top_container = ttk.Frame(main_paned)
        main_paned.add(top_container, weight=2)
        
        canvas = tk.Canvas(top_container, highlightthickness=0)
        v_scroll = ttk.Scrollbar(top_container, orient="vertical", command=canvas.yview)
        form_scrollable = ttk.Frame(canvas)
        
        form_scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        _canvas_win = canvas.create_window((0, 0), window=form_scrollable, anchor="nw")
        canvas.configure(yscrollcommand=v_scroll.set)
        
        # [FIX] 우측 빈 공간 제거: Canvas 너비 변경 시 내부 프레임 너비를 동기화
        def _on_canvas_resize(event):
            canvas.itemconfigure(_canvas_win, width=event.width)
        canvas.bind("<Configure>", _on_canvas_resize)
        
        # Scroll binding is handled globally in MaterialManager.__init__
        
        canvas.pack(side="left", fill="both", expand=True)
        v_scroll.pack(side="right", fill="y")
        
        # 1. Main Budget Form
        form_frame = ttk.LabelFrame(form_scrollable, text="실행예산 입력/수정", padding=10)
        form_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        # Grid layout for form fields
        rows = [
            ("현장명:", "cb_budget_site"),
            ("도급액 (Revenue):", "ent_budget_revenue"),
            ("검사단가 (UnitPrice):", "ent_budget_unit_price"),
            ("실행 노무비 (Labor):", "ent_budget_labor"),
            ("실행 재료비 (Material):", "ent_budget_material"),
            ("실행 경비 (Expense):", "ent_budget_expense"),
            ("실행 외주비 (Outsource):", "ent_budget_outsource"),
            ("영업이익 (Profit):", "ent_budget_profit"),
            ("비고:", "ent_budget_note")
        ]
        
        self.budget_widgets = {}
        for i, (label_text, attr_name) in enumerate(rows):
            ttk.Label(form_frame, text=label_text).grid(row=i//4*2, column=i%4, sticky='w', padx=5, pady=2)
            if attr_name == "cb_budget_site":
                widget = ttk.Combobox(form_frame, width=30)
                widget['values'] = self.sites
            else:
                widget = ttk.Entry(form_frame, width=32)
                if attr_name == "ent_budget_labor":
                    # [STABILITY] Make labor cost read-only if it's auto-calculated by the widget?
                    # No, let the user override if needed, but the widget will update it.
                    pass
                
            widget.grid(row=i//4*2+1, column=i%4, sticky='ew', padx=5, pady=2)
            setattr(self, attr_name, widget)
            self.budget_widgets[attr_name] = widget

        def _on_budget_site_selected(e):
            # [FIX] 현장명을 가장 먼저 캡처 - focus_set() 호출 전에 해야 값이 유지됨
            selected_site = self.cb_budget_site.get().strip()
            # SuggestionWindow(자동완성 창) 강제 닫기
            if hasattr(self, '_suggestion_windows') and self.cb_budget_site in self._suggestion_windows:
                try:
                    self._suggestion_windows[self.cb_budget_site].hide()
                except Exception:
                    pass
            # 폼 로드 (현장명 먼저 캡처했으므로 안전)
            self._load_budget_to_form(selected_site, silent=True)
            # 포커스 이동은 폼 로드 완료 후에 (after로 지연)
            self.root.after(50, self.ent_budget_revenue.focus_set)
        self.cb_budget_site.bind('<<ComboboxSelected>>', _on_budget_site_selected)

        # 2. Detailed Labor Cost Widget
        labor_detail_frame = ttk.LabelFrame(form_scrollable, text="인건비 상세 (Labor Cost Detail)", padding=10)
        labor_detail_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        def on_labor_change(total):
            self.ent_budget_labor.delete(0, tk.END)
            self.ent_budget_labor.insert(0, f"{total:,.0f}")
            self._update_budget_kpis()
            
        self.labor_detail_widget = LaborCostDetailWidget(labor_detail_frame, on_change_callback=on_labor_change)
        self.labor_detail_widget.pack(fill='x', expand=True)

        # 3. Detailed Material Cost Widget
        material_detail_frame = ttk.LabelFrame(form_scrollable, text="재료비 상세 (Material Cost Detail)", padding=10)
        material_detail_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        def on_material_change(total):
            self.ent_budget_material.delete(0, tk.END)
            self.ent_budget_material.insert(0, f"{total:,.0f}")
            self._update_budget_kpis()
            
        self.material_detail_widget = MaterialCostDetailWidget(material_detail_frame, on_change_callback=on_material_change)
        self.material_detail_widget.pack(fill='x', expand=True)

        # 4. Detailed Expense & Profit Widget
        expense_detail_frame = ttk.LabelFrame(form_scrollable, text="경비 및 이익 상세 (Expense & Profit Detail)", padding=10)
        expense_detail_frame.pack(fill='x', pady=(0, 10), padx=5)
        
        def on_expense_change(exp_total, outsource_total, op_profit):
            self.ent_budget_expense.delete(0, tk.END)
            self.ent_budget_expense.insert(0, f"{exp_total:,.0f}")
            self.ent_budget_outsource.delete(0, tk.END)
            self.ent_budget_outsource.insert(0, f"{outsource_total:,.0f}")
            # [STABILITY] Update profit field if it exists
            if hasattr(self, 'ent_budget_profit'):
                # Calculate margin
                rev = get_rev()
                margin = (op_profit / rev * 100) if rev > 0 else 0.0
                self.ent_budget_profit.delete(0, tk.END)
                self.ent_budget_profit.insert(0, f"{op_profit:,.0f} ({margin:.1f}%)")
            self._update_budget_kpis()

        def get_lab():
            try: return float(self.ent_budget_labor.get().replace(',', '') or 0)
            except: return 0.0
        def get_mat():
            try: return float(self.ent_budget_material.get().replace(',', '') or 0)
            except: return 0.0
        def get_rev():
            try: 
                rev = float(self.ent_budget_revenue.get().replace(',', '') or 0)
                if rev == 0 and hasattr(self, 'ent_budget_unit_price'):
                    rev = float(self.ent_budget_unit_price.get().replace(',', '') or 0)
                return rev
            except: return 0.0
            
        self.expense_detail_widget = ExpenseProfitDetailWidget(
            expense_detail_frame, 
            on_change_callback=on_expense_change,
            get_labor_total_func=get_lab,
            get_material_total_func=get_mat,
            get_revenue_func=get_rev
        )
        self.expense_detail_widget.pack(fill='x', expand=True)
        
        # [NEW] 도급액 실시간 입력에 따른 이익금액 재계산 바인딩
        if hasattr(self, 'ent_budget_revenue'):
            self.ent_budget_revenue.bind('<KeyRelease>', lambda e: self.expense_detail_widget.calculate_all())

        # [MOVED to top_btn_frame]

        # ===============================================================
        # --- 하단: 현장별 일일사용량 조회 프레임 ---
        # ===============================================================
        bottom_container = ttk.Frame(main_paned)
        main_paned.add(bottom_container, weight=3)

        # 필터 영역
        bottom_filter = ttk.Frame(bottom_container)
        bottom_filter.pack(fill='x', pady=(5, 2))

        ttk.Label(bottom_filter, text="현장별 실적 조회", font=('Malgun Gothic', 10, 'bold')).pack(side='left', padx=(0, 15))

        ttk.Label(bottom_filter, text="현장:").pack(side='left', padx=5)
        self.cb_budget_view_site = ttk.Combobox(bottom_filter, width=20, state='readonly')
        self.cb_budget_view_site['values'] = self.sites
        self.cb_budget_view_site.pack(side='left', padx=5)
        self.cb_budget_view_site.bind('<<ComboboxSelected>>', lambda e: self.update_budget_site_view())

        ttk.Label(bottom_filter, text="시작일:").pack(side='left', padx=(10, 2))
        self.budget_view_start = DateEntry(bottom_filter, width=12, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        self.budget_view_start.pack(side='left', padx=2)
        today = datetime.datetime.now()
        self.budget_view_start.set_date(today.replace(day=1))

        ttk.Label(bottom_filter, text="종료일:").pack(side='left', padx=(5, 2))
        self.budget_view_end = DateEntry(bottom_filter, width=12, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        self.budget_view_end.pack(side='left', padx=2)
        self.budget_view_end.set_date(today)

        ttk.Button(bottom_filter, text="조회", command=self.update_budget_site_view).pack(side='left', padx=8)
        ttk.Button(bottom_filter, text="예산 수정/불러오기",
                   command=lambda: self._load_budget_to_form(self.cb_budget_view_site.get())).pack(side='left', padx=4)
        ttk.Button(bottom_filter, text="엑셀 내보내기", command=self.export_budget_sales_status).pack(side='right', padx=10)

        # KPI 요약 행 제거 (예산입력 5번 영업이익률로 대체)
        # Treeview (일일사용량 코럼)
        tree_outer = ttk.Frame(bottom_container)
        tree_outer.pack(fill='both', expand=True)

        bv_cols = ('Date', 'Site', '장비명', '검사방법', '검사량', '단가', '검사단가', '출장비', '일식', 'OT합계', '자재사용량', '자재단가', '합계', '비고')
        bv_widths = (90, 120, 100, 80, 70, 80, 90, 80, 70, 80, 80, 90, 100, 150)
        bv_heads  = ('날짜', '현장', '장비', '검사방법', '수량', '단가', '검사단가', '출장비', '일식', 'OT합계', '자재사용량', '자재단가', '합계', '비고')

        vsb = ttk.Scrollbar(tree_outer, orient='vertical')
        hsb = ttk.Scrollbar(tree_outer, orient='horizontal')
        self.budget_view_tree = ttk.Treeview(tree_outer, columns=bv_cols, show='headings',
                                              yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        for col, head, w in zip(bv_cols, bv_heads, bv_widths):
            self.budget_view_tree.heading(col, text=head,
                command=lambda c=col: self.treeview_sort_column(self.budget_view_tree, c, False))
            self.budget_view_tree.column(col, width=w, minwidth=40, anchor='center')

        vsb.config(command=self.budget_view_tree.yview)
        hsb.config(command=self.budget_view_tree.xview)
        self.budget_view_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        tree_outer.grid_rowconfigure(0, weight=1)
        tree_outer.grid_columnconfigure(0, weight=1)

        # ── ESC: 현장별 탭 내 모든 입력 위젯에서 포커스 해제 ──────────────
        def _esc_to_root(e=None):
            self.root.focus_set()
            return "break"

        def _bind_esc_recursive(widget):
            try:
                cls = widget.winfo_class()
                # Toplevel(달력 팝업 등) 내부는 절대 순회하지 않음
                # → 순회만으로도 winfo_children() 가 Toplevel을 화면에 표시시킬 수 있음
                if cls == 'Toplevel':
                    return
                if cls in ('TEntry', 'Entry', 'TCombobox'):
                    widget.bind('<Escape>', _esc_to_root, add='+')
            except Exception:
                pass
            for child in widget.winfo_children():
                _bind_esc_recursive(child)

        # after_idle: 모든 위젯 생성 완료 후 바인딩
        self.root.after_idle(lambda: _bind_esc_recursive(self.tab_daily_usage))

    def update_budget_view(self):
        """Refresh the budget treeview - no-op (UI triggered)"""
        pass

    def update_budget_site_view(self):
        """현장별 일일사용량 데이터를 하단 Treeview에 표시하고 KPI를 갱신한다."""
        if not hasattr(self, 'budget_view_tree'):
            return

        for item in self.budget_view_tree.get_children():
            self.budget_view_tree.delete(item)

        site = self.cb_budget_view_site.get().strip()
        if not site:
            return

        if self.daily_usage_df.empty:
            return

        # --- 날짜 필터 ---
        start_date = self.budget_view_start.get_date()
        end_date   = self.budget_view_end.get_date()
        df = self.daily_usage_df.copy()
        df['_date'] = pd.to_datetime(df['Date'], errors='coerce').dt.date
        mask = (df['_date'] >= start_date) & (df['_date'] <= end_date) & (df['Site'] == site)
        df = df[mask].copy()

        # --- 자재 원가 맵 ---
        mat_id_cost_map = {}
        if 'MaterialID' in self.materials_df.columns and '원가' in self.materials_df.columns:
            mat_id_cost_map = self.materials_df.set_index('MaterialID')['원가'].fillna(0).to_dict()

        def _f(val):
            if pd.isna(val) or str(val).lower() == 'nan': return 0.0
            try: 
                s = str(val).replace(',', '').strip()
                if not s: return 0.0
                num = float(s)
                return 0.0 if np.isnan(num) else num
            except: return 0.0

        total_income = 0.0   # 검사비 + OT 합계
        total_expense = 0.0  # 출장비 + 일식
        total_net_revenue = 0.0 # 검사비 합계
        total_sum     = 0.0
        total_mat_cost = 0.0 # 자재 실적 사용료 합계

        import re as _re
        for _, row in df.iterrows():
            # [FIX] OT 컬럼에서 금액 추출 로직 강화 (NaN 방지)
            parsed_ots = []
            for i in range(1, 11):
                raw_ot = row.get('OT' if i == 1 else f'OT{i}', 0)
                if pd.isna(raw_ot) or not str(raw_ot).strip() or str(raw_ot).lower() == 'nan':
                    parsed_ots.append(0.0)
                    continue
                
                v_str = str(raw_ot).strip()
                if '(' in v_str and '원)' in v_str:
                    try:
                        amt_str = v_str.split('(')[1].split('원')[0].replace(',', '').strip()
                        parsed_ots.append(_f(amt_str))
                    except: parsed_ots.append(0.0)
                else:
                    parsed_ots.append(_f(v_str))
            
            ot_sum = sum(parsed_ots)

            net    = _f(row.get('검사비', 0))
            travel = _f(row.get('출장비', 0))
            meal   = _f(row.get('일식', 0))
            qty    = _f(row.get('검사량', 0))
            price  = _f(row.get('단가', 0))

            # 자재원가
            test_method = str(row.get('검사방법', '')).upper()
            if 'RT' in test_method:
                usage = _f(row.get('FilmCount', 0))
            else:
                usage = _f(row.get('Usage', 0))
            unit_cost = mat_id_cost_map.get(row.get('MaterialID'), 0)
            mat_cost = usage * float(unit_cost)

            income  = net + ot_sum
            expense = travel + meal
            row_sum = income + expense

            total_income   += income
            total_expense  += expense
            total_net_revenue += net
            total_sum      += row_sum
            total_mat_cost += mat_cost

            self.budget_view_tree.insert('', tk.END, values=(
                str(row.get('Date', '')),
                str(row.get('Site', '')),
                str(row.get('장비명', '')),
                str(row.get('검사방법', '')),
                f"{qty:g}",
                f"{price:,.0f}",
                f"{net:,.0f}",
                f"{travel:,.0f}",
                f"{meal:,.0f}",
                f"{ot_sum:,.0f}",
                f"{usage:g}",
                f"{unit_cost:,.0f}",
                f"{row_sum:,.0f}",
                str(row.get('비고', row.get('Note', ''))).replace('nan', '')
            ))

    def _update_budget_kpis(self):
        """Update the top KPI summary labels based on current budget form values"""
        try:
            def _get_val(widget):
                try: return float(str(widget.get()).replace(',', '').split(' ')[0] or 0)
                except: return 0.0
                
            rev = _get_val(self.ent_budget_revenue)
            if rev == 0 and hasattr(self, 'ent_budget_unit_price'):
                rev = _get_val(self.ent_budget_unit_price) # Fallback to unit price if revenue empty
                
            lab = _get_val(self.ent_budget_labor)
            mat = _get_val(self.ent_budget_material)
            exp = _get_val(self.ent_budget_expense)
            out = _get_val(self.ent_budget_outsource)
            
            total_cost = lab + mat + exp + out
            profit = rev - total_cost
            margin = (profit / rev * 100) if rev > 0 else 0.0
            
            if hasattr(self, 'lbl_kpi_rev'):
                self.lbl_kpi_rev.config(text=f"{rev:,.0f}원")
                self.lbl_kpi_cost.config(text=f"{total_cost:,.0f}원")
                self.lbl_kpi_profit.config(text=f"{profit:,.0f}원", foreground="#ef4444" if profit < 0 else "#10b981")
                self.lbl_kpi_margin.config(text=f"{margin:.1f}%", foreground="#ef4444" if margin < 0 else "#10b981")
        except:
            pass

    def _load_budget_to_form(self, site, silent=False):
        """선택한 현장의 예산서를 상단 입력 폼에 로드한다."""
        # [FIX] 자동완성 SuggestionWindow 강제 닫기
        if hasattr(self, 'cb_budget_site') and hasattr(self.cb_budget_site, '_suggestion_win'):
            try:
                self.cb_budget_site._suggestion_win.hide()
            except Exception:
                pass
        # [NEW] Clear current form before loading new site data to prevent residual data
        self.clear_budget_form()

        # [FIX] clear_budget_form()이 cb_budget_site를 지우므로 즉시 복원
        # budget_df에 해당 현장이 없더라도 콤보박스에 선택한 현장명은 유지해야 함
        if site and hasattr(self, 'cb_budget_site'):
            self.cb_budget_site.set(site)

        if not site:
            if not silent: messagebox.showwarning("현장 미선택", "현장을 먼저 선택해주세요.")
            return
        if not hasattr(self, 'budget_df') or self.budget_df.empty:
            if not silent: messagebox.showinfo("데이터 없음", f"'{site}' 현장의 저장된 예산서가 없습니다.")
            return
        match = self.budget_df[self.budget_df['Site'] == site]
        if match.empty:
            if not silent: messagebox.showinfo("데이터 없음", f"'{site}' 현장의 저장된 예산서가 없습니다.")
            return

        row = match.iloc[0]
        self.cb_budget_site.set(site)
        for attr, col in [('ent_budget_revenue',   'Revenue'),
                          ('ent_budget_unit_price','UnitPrice'),
                          ('ent_budget_labor',     'LaborCost'),
                          ('ent_budget_material',  'MaterialCost'),
                          ('ent_budget_expense',   'Expense'),
                          ('ent_budget_outsource', 'OutsourceCost'),
                          ('ent_budget_profit',    'Profit'),
                          ('ent_budget_note',      'Note')]:
            w = getattr(self, attr, None)
            if w:
                w.delete(0, tk.END)
                val = row.get(col, '')
                try:
                    w.insert(0, f"{float(val):,.0f}" if col != 'Note' and val else str(val))
                except:
                    w.insert(0, str(val))

        import json as _json
        for json_col, widget_attr in [('LaborDetail',   'labor_detail_widget'),
                                       ('MaterialDetail','material_detail_widget'),
                                       ('ExpenseDetail', 'expense_detail_widget')]:
            jdata = row.get(json_col, '')
            widget = getattr(self, widget_attr, None)
            if widget and jdata and str(jdata).strip():
                try:
                    widget.set_data(_json.loads(jdata))
                except:
                    widget.reset()
            elif widget:
                # [NEW] Clear data if no saved detail exists for this site
                widget.reset()

        # [FIX] set_data() → calculate_all() → on_change_callback() 순서로 금액 필드가
        # 상세 항목 합계로 덮어씌워지는 문제 방지.
        # 상세 위젯 로드 완료 후 저장된 원본 금액 값을 명시적으로 재설정한다.
        for attr, col in [('ent_budget_revenue',   'Revenue'),
                          ('ent_budget_unit_price','UnitPrice'),
                          ('ent_budget_labor',     'LaborCost'),
                          ('ent_budget_material',  'MaterialCost'),
                          ('ent_budget_expense',   'Expense'),
                          ('ent_budget_outsource', 'OutsourceCost'),
                          ('ent_budget_profit',    'Profit')]:
            w = getattr(self, attr, None)
            if w:
                val = row.get(col, '')
                try:
                    display = f"{float(val):,.0f}" if val else '0'
                    if attr == 'ent_budget_profit':
                        rev_val = row.get('Revenue', 0)
                        try: rev_float = float(rev_val)
                        except: rev_float = 0.0
                        prof_float = float(val) if val else 0.0
                        margin = (prof_float / rev_float * 100) if rev_float > 0 else 0.0
                        display = f"{prof_float:,.0f} ({margin:.1f}%)"
                except:
                    display = '0'
                w.delete(0, tk.END)
                w.insert(0, display)

        if not silent: messagebox.showinfo("로드 완료", f"'{site}' 현장의 예산서를 불러왔습니다.")
        # [STABILITY] Update KPI summary after load
        self._update_budget_kpis()

    def fill_budget_from_actuals(self):
        """하단 필터(현장, 기간) 기준으로 실제 실적 데이터를 집계하여 상단 예산 폼에 자동 입력한다."""
        site = self.cb_budget_view_site.get().strip()
        if not site:
            messagebox.showwarning("현장 미선택", "하단 '현장별 실적 조회' 영역에서 현장을 먼저 선택해주세요.")
            return

        start_ts = pd.Timestamp(self.budget_view_start.get_date())
        end_ts = pd.Timestamp(self.budget_view_end.get_date())

        if self.daily_usage_df.empty:
            messagebox.showinfo("데이터 없음", "집계할 실적 데이터가 없습니다.")
            return

        # 1. 필터링
        df = self.daily_usage_df.copy()
        df['_date_ts'] = pd.to_datetime(df['Date'], errors='coerce')
        mask = (df['_date_ts'] >= start_ts) & (df['_date_ts'] <= end_ts) & (df['Site'] == site)
        df = df[mask].copy()

        if df.empty:
            s_date = self.budget_view_start.get_date()
            e_date = self.budget_view_end.get_date()
            messagebox.showinfo("데이터 없음", f"선택한 기간 ({s_date} ~ {e_date}) 동안 '{site}' 현장의 실적이 없습니다.")
            return

        # 2. 집계 초기화
        def _f(val):
            if pd.isna(val) or str(val).lower() == 'nan': return 0.0
            try: 
                s = str(val).replace(',', '').strip()
                if not s: return 0.0
                num = float(s)
                return 0.0 if np.isnan(num) else num
            except: return 0.0

        total_net_revenue = 0.0 # 검사비
        total_travel = 0.0
        total_meal = 0.0
        
        # --- Labor Aggregation by Rank ---
        ranks = ["이사", "부장", "차장", "과장", "대리", "계장", "주임", "기사"]
        # Track unique dates for each worker to calculate active days correctly
        # Structure: {rank: {worker_name: set(date_strings)}}
        rank_labor_dates = {r: {} for r in ranks}
        
        # --- Special Work (OT) Classification ---
        special_types = ["연장근무", "야간근무", "휴일근무"]
        ot_data = {t: {'names': set(), 'hours': 0} for t in special_types}

        import re as _re
        # Updated pattern to support [Rank], Rank Name, or Name alone with mapping
        # Supports: "[부장] 주진철", "부장 주진철", "주진철"
        rank_pattern = _re.compile(r"\[?(이사|부장|차장|과장|대리|계장|주임|기사)\]?\s*(.*)")
        
        # Default mapping for core workers if rank is missing
        worker_rank_map = {
            "주진철": "부장", "우명광": "대리", "김진환": "주임", "장승대": "계장",
            "김성렬": "주임", "박광복": "부장", "주영광": "과장", "이경재": "주임",
            "황조현": "주임", "김춘호": "차장", "박원준": "과장", "이봉주": "부장",
            "주동기": "부장", "김성현": "대리"
        }

        # 상세 재료비 집계를 위한 맵
        material_usage_sums = [0.0] * 10
        mat_map = {
            '세척': 0, '침투': 1, '현상': 2, '자분': 3, '흑색': 4,
            '필름': 5, '글리세': 6, '현상액': 7, '정착액': 8, '수적': 9
        }

        # 자재 마스터 정보
        mat_id_name_map = {}
        mat_id_cost_map = {}
        if not self.materials_df.empty:
            mat_id_name_map = self.materials_df.set_index('MaterialID')['품목명'].fillna('').to_dict()
            mat_id_cost_map = self.materials_df.set_index('MaterialID')['원가'].fillna(0).to_dict()

        total_mat_cost = 0.0
        rt_inspection_count = 0  # Counter for RT occurrences
        # [NEW] Use unique days instead of row count to avoid inflating days for multiple records
        total_days_count = df['_date_ts'].dt.date.nunique() if not df.empty else 0
        
        # [NEW] Track unique dates for vehicles
        starex_dates = set()
        toptruck_dates = set()

        for _, row in df.iterrows():
            # Check weekday
            date_val = _re.sub(r'\s.*', '', str(row.get('Date', '')))
            try:
                dt_obj = pd.to_datetime(date_val)
                is_weekday = dt_obj.weekday() < 5 # 0-4: Mon-Fri
            except:
                is_weekday = True

            total_net_revenue += _f(row.get('검사비', 0))
            total_travel += _f(row.get('출장비', 0))
            total_meal += _f(row.get('일식', 0))
            
            # --- Vehicle Tracking ---
            car_info = f"{row.get('차량번호', '')} {row.get('차량비고', '')} {row.get('장비명', '')}".strip()
            
            try:
                dt_obj = pd.to_datetime(date_val)
                date_key = dt_obj.date()
            except:
                date_key = date_val
                
            if '스타렉스' in car_info:
                starex_dates.add(date_key)
            if '탑차' in car_info:
                toptruck_dates.add(date_key)

            # --- Labor & OT Parsing ---
            shift = str(row.get('Shift', '주간')).strip()
            for i in range(1, 11):
                suffix = '' if i == 1 else str(i)
                raw_worker = str(row.get(f'User{suffix}', '')).strip()
                if not raw_worker or raw_worker.lower() == 'nan':
                    continue
                
                # Rank Aggregation
                match = rank_pattern.search(raw_worker)
                worker_name_only = raw_worker
                rank = None
                
                if match:
                    # Case 1: Rank identified via pattern (e.g., "[부장] 주진철" or "부장 주진철")
                    if match.group(1) in ranks:
                        rank = match.group(1)
                        worker_name_only = match.group(2).strip()
                        if not worker_name_only: # Only rank was present
                             worker_name_only = rank
                
                # Case 2: Fallback to mapping if rank not identified from the string
                if not rank:
                    # Clean potential shift prefixes if they exist (though usually cleaned in records)
                    clean_name = _re.sub(r'\(.*?\)', '', raw_worker).strip()
                    rank = worker_rank_map.get(clean_name)
                    worker_name_only = clean_name
                
                if rank and rank in rank_labor_dates:
                    if worker_name_only not in rank_labor_dates[rank]:
                        rank_labor_dates[rank][worker_name_only] = set()
                    try:
                        # Normalize date key for uniqueness
                        dt_obj = pd.to_datetime(date_val)
                        d_key = dt_obj.date()
                    except:
                        d_key = date_val
                    rank_labor_dates[rank][worker_name_only].add(d_key)
                
                # [FIX] WorkTime 컬럼에서 분류(주간/야간/휴일)와 시간 파싱
                # 형식: "(야간) 4h" or "(주야간) 2.5h" 또는 "8h" (구형식)
                wt_col = f'WorkTime{suffix}'
                wt_val = str(row.get(wt_col, '')).strip()
                
                # WorkTime에서 분류 파싱
                wt_shift = shift  # 기본값: 레코드의 Shift 컬럼
                wt_hours = 0.0
                if wt_val and wt_val.lower() != 'nan':
                    # Parse shift type if present, e.g. "(야간)", "(연장)"
                    shift_match = _re.search(r'\(?(주간|야간|주야간|휴일|연장)\)?', wt_val)
                    if shift_match:
                        wt_shift = shift_match.group(1)
                    
                    # Compute duration calculation
                    marker_pattern = MARKER_PATTERN
                    clean_wt = marker_pattern.sub('', wt_val).replace('(연장)', '').replace('(주야간)', '').replace('(야간)', '').replace('(휴일)', '').replace('(주간)', '').strip()
                    sep_match = _re.search(r'[:\d]\s*([~-])\s*[:\d]', clean_wt)
                    
                    if sep_match:
                        try:
                            sep = sep_match.group(1)
                            st_str, en_str = clean_wt.split(sep)
                            sh, sm = map(int, st_str.split(':'))
                            eh, em = map(int, en_str.split(':'))
                            start_f = sh + sm/60.0
                            end_f = eh + em/60.0
                            if end_f < start_f: end_f += 24
                            wt_hours = end_f - start_f
                        except:
                            pass
                    else:
                        # Fallback: Just a number like "8h" or "8"
                        h_match = _re.search(r'([\d.]+)', clean_wt)
                        if h_match:
                            try: wt_hours = float(h_match.group(1))
                            except: pass
                
                # [FIX] 분류별 OT 상세 집계
                if wt_hours > 0:
                    # 명시적으로 '휴일'이라고 쓰여있거나, 주말인데 기본 교대근무가 아닌 특근성 출근인 경우 (이번엔 단순히 텍스트 명시에 더 의존)
                    if '휴일' in wt_shift:
                        ot_data['휴일근무']['hours'] += wt_hours
                        ot_data['휴일근무']['names'].add(worker_name_only)
                    else:
                        # 평일/주말 공통: 연장이나 야간 명시가 있거나 시간대 기반 계산
                        if sep_match:
                            try:
                                sep = sep_match.group(1)
                                st_str, en_str = clean_wt.split(sep)
                                sh, sm = map(int, st_str.split(':'))
                                eh, em = map(int, en_str.split(':'))
                                start_f = sh + sm/60.0
                                end_f = eh + em/60.0
                                if end_f < start_f: end_f += 24
                                
                                # 현장 기본 퇴근 시간 (17시 기준 혹은 18시 등)
                                # 전체 시간이 주어질 때 명확히 구분 (기본근무: ~18, 연장근무: 18~22, 야간근무: 22~)
                                if end_f > 18:
                                    ot_start = max(start_f, 18.0)
                                    # 18~22: 연장근무
                                    h_over = max(0, min(end_f, 22.0) - ot_start)
                                    # 22~24 : 야간근무
                                    h_night = max(0, min(end_f, 24.0) - max(ot_start, 22.0))
                                    # 24~ : 심야/익일 근무
                                    h_dawn = max(0, end_f - max(ot_start, 24.0))
                                    
                                    if h_over > 0:
                                        ot_data['연장근무']['hours'] += h_over
                                        ot_data['연장근무']['names'].add(worker_name_only)
                                    if h_night > 0:
                                        ot_data['야간근무']['hours'] += h_night
                                        ot_data['야간근무']['names'].add(worker_name_only)
                                    if h_dawn > 0:
                                        try:
                                            # 날짜 확인 (금요일 밤 -> 토요일 새벽인지)
                                            entry_date = pd.to_datetime(row.get('Date'))
                                            if entry_date.weekday() == 4: # 4 = 금요일
                                                ot_data['휴일근무']['hours'] += h_dawn
                                                ot_data['휴일근무']['names'].add(worker_name_only)
                                            else:
                                                ot_data['야간근무']['hours'] += h_dawn
                                                ot_data['야간근무']['names'].add(worker_name_only)
                                        except:
                                            ot_data['야간근무']['hours'] += h_dawn
                                            ot_data['야간근무']['names'].add(worker_name_only)
                            except: pass
                        else:
                            # 시간 범위가 주어지지 않은 숫자(예: 4h, 10h) 처리 로직
                            # 명시적으로 (연장) 태그가 있는 경우
                            if '연장' in wt_shift:
                                ot_data['연장근무']['hours'] += wt_hours
                                ot_data['연장근무']['names'].add(worker_name_only)
                            # 명시적 태그가 없으나 8시간을 초과하는 경우
                            elif wt_hours > 8:
                                h_extra = wt_hours - 8
                                if '야간' in wt_shift or '주야간' in wt_shift:
                                    ot_data['야간근무']['hours'] += h_extra
                                    ot_data['야간근무']['names'].add(worker_name_only)
                                else:
                                    ot_data['연장근무']['hours'] += h_extra
                                    ot_data['연장근무']['names'].add(worker_name_only)


            # --- 자재 사용량 집계 ---
            m_id = row.get('MaterialID', '')
            m_name = str(mat_id_name_map.get(m_id, '')).strip()
            test_method = str(row.get('검사방법', '')).upper()
            
            usage = 0.0
            if 'RT' in test_method:
                rt_inspection_count += 1  # [NEW] Check for RT
                usage = _f(row.get('FilmCount', 0))
                material_usage_sums[5] += usage
            else:
                usage = _f(row.get('Usage', 0))
                for keyword, idx in mat_map.items():
                    if keyword in m_name:
                        if keyword == '현상' and ('액' in m_name or 'PT' not in test_method):
                             if 'PT' in test_method: 
                                 material_usage_sums[2] += usage
                             else:
                                 material_usage_sums[7] += usage
                             break
                        material_usage_sums[idx] += usage
                        break
            
            total_mat_cost += usage * float(mat_id_cost_map.get(m_id, 0))

        # 3. 폼에 입력 및 [5. 영업이익] '실행(실적)' 줄에 표시 결과 연동
        # Update Special Unit Prices if LaborCostDetailWidget is available to compute correct revenue
        ot_revenue = 0.0
        if hasattr(self, 'labor_detail_widget'):
            ot_revenue = sum(ot_data[t]['hours'] * _f(self.labor_detail_widget.entries[t]['unit_price'].get()) 
                            for t in special_types if t in self.labor_detail_widget.entries)

        self.cb_budget_site.set(site)
        self.ent_budget_revenue.delete(0, tk.END)
        # 사용 요청: 도급액은 비워두고 입력 시 반영되도록 함. 
        
        # [NEW] 실적 데이터의 검사비 총액을 '검사단가' 칸에 기본 입력
        if hasattr(self, 'ent_budget_unit_price'):
            self.ent_budget_unit_price.delete(0, tk.END)
            self.ent_budget_unit_price.insert(0, f"{total_net_revenue:,.0f}")
        
        self.ent_budget_material.delete(0, tk.END)
        self.ent_budget_material.insert(0, f"{total_mat_cost:,.0f}")
        
        self.ent_budget_expense.delete(0, tk.END)
        self.ent_budget_expense.insert(0, f"{(total_travel + total_meal):,.0f}")

        # 4. 인건비 상세 위젯 업데이트
        if hasattr(self, 'labor_detail_widget'):
            labor_data = self.labor_detail_widget.get_data()
            # Regular
            for rank in ranks:
                if rank in labor_data:
                    # Number of unique workers who logged time in this rank
                    u_count = len(rank_labor_dates[rank])
                    # Total number of unique working days summed across all workers in this rank
                    t_days = sum(len(dates) for dates in rank_labor_dates[rank].values())
                    if u_count > 0:
                        avg_d = t_days / u_count
                        labor_data[rank]['personnel'] = f"{u_count:g}"
                        labor_data[rank]['period'] = f"{avg_d:g}"
                    else:
                        labor_data[rank]['personnel'] = ""
                        labor_data[rank]['period'] = ""
            # Special (OT)
            for stype in special_types:
                if stype in labor_data:
                    u_count = len(ot_data[stype]['names'])
                    t_hours = ot_data[stype]['hours']
                    if u_count > 0:
                        avg_h = t_hours / u_count
                        labor_data[stype]['personnel'] = f"{u_count:g}"
                        labor_data[stype]['period'] = f"{avg_h:g}"
                    else:
                        labor_data[stype]['personnel'] = ""
                        labor_data[stype]['period'] = ""
            self.labor_detail_widget.set_data(labor_data)

        # 5. 재료비 상세 위젯 업데이트
        if hasattr(self, 'material_detail_widget'):
            current_mat_data = self.material_detail_widget.get_data()
            for i, qty in enumerate(material_usage_sums):
                if i < len(current_mat_data):
                    current_mat_data[i]['qty'] = f"{qty:g}" if qty > 0 else ""
            self.material_detail_widget.set_data(current_mat_data)
        
        # [NEW] Immediate KPI update after data fill
        self._update_budget_kpis()
            
        # 6. 경비 상세 위젯 업데이트 (차량유지비, 소모품비, 복리후생비, Se-175 수량, 감가상각비 차량 사용일수 자동 할당)
        if hasattr(self, 'expense_detail_widget'):
            self.expense_detail_widget.calculate_all()
            
            # [STABILITY] Update KPI summary after loading
            self._update_budget_kpis()
            current_exp_data = self.expense_detail_widget.get_data()
            for row_data in current_exp_data.get('site_expense', []):
                cat = row_data.get('cat')
                if cat in ['차량유지비', '소모품비', '복리후생비']:
                    row_data['qty'] = f"{total_days_count:g}"
                elif cat == 'Se-175':
                    row_data['qty'] = f"{rt_inspection_count:g}"
                    
            for row_data in current_exp_data.get('depreciation', []):
                item = row_data.get('item', '')
                if '스타렉스' in item:
                    row_data['days'] = f"{len(starex_dates):g}" if len(starex_dates) > 0 else ""
                elif '탑차' in item:
                    row_data['days'] = f"{len(toptruck_dates):g}" if len(toptruck_dates) > 0 else ""
                    
            self.expense_detail_widget.set_data(current_exp_data)

        # [NEW] Calculate profit/margin for the summary message
        try:
            rev = float(self.ent_budget_unit_price.get().replace(',', '') or 0)
            lab = float(self.ent_budget_labor.get().replace(',', '') or 0)
            mat = float(self.ent_budget_material.get().replace(',', '') or 0)
            exp = float(self.ent_budget_expense.get().replace(',', '') or 0)
            out = float(self.ent_budget_outsource.get().replace(',', '') or 0)
            cost = lab + mat + exp + out
            profit = rev - cost
            margin = (profit / rev * 100) if rev > 0 else 0.0
            profit_str = f"\n\n▶ 집계 결과\n- 도급액(검사비): {rev:,.0f}원\n- 실행원가: {cost:,.0f}원\n- 영업이익: {profit:,.0f}원 ({margin:.1f}%)"
        except:
            profit_str = ""

        messagebox.showinfo("조회 완료", 
                            f"'{site}' 현장의 실적 데이터를 집계하여 [5. 영업이익] 탭에 표시하였습니다.\n"
                            f"(기간: {start_ts.strftime('%Y-%m-%d')} ~ {end_ts.strftime('%Y-%m-%d')}){profit_str}")

    def save_budget_entry(self):
        """Save or update budget entry"""
        site = self.cb_budget_site.get().strip()
        if not site:
            messagebox.showwarning("입력 오류", "현장명을 입력해주세요.")
            return
            
        try:
            rev = float(str(self.ent_budget_revenue.get()).replace(',', '') or 0)
            unit_price_val = str(getattr(self, 'ent_budget_unit_price', tk.Entry()).get() if hasattr(self, 'ent_budget_unit_price') else '').replace(',', '').strip()
            unit_price = float(unit_price_val) if unit_price_val else 0.0
            lab = float(str(self.ent_budget_labor.get()).replace(',', '') or 0)
            mat = float(str(self.ent_budget_material.get()).replace(',', '') or 0)
            exp = float(str(self.ent_budget_expense.get()).replace(',', '') or 0)
            out = float(str(self.ent_budget_outsource.get()).replace(',', '') or 0)
            note = self.ent_budget_note.get().strip()
            profit = rev - (lab + mat + exp + out)
        except ValueError:
            messagebox.showerror("입력 오류", "금액은 숫자여야 합니다.")
            return

        new_data = {
            'Site': site,
            'Revenue': rev,
            'UnitPrice': unit_price,
            'LaborCost': lab,
            'MaterialCost': mat,
            'Expense': exp,
            'OutsourceCost': out,
            'Profit': profit,
            'Note': note,
            'LaborDetail': json.dumps(self.labor_detail_widget.get_data()),
            'MaterialDetail': json.dumps(self.material_detail_widget.get_data()),
            'ExpenseDetail': json.dumps(self.expense_detail_widget.get_data())
        }

        if not self.budget_df.empty and site in self.budget_df['Site'].values:
            idx = self.budget_df[self.budget_df['Site'] == site].index[0]
            for key, val in new_data.items():
                self.budget_df.at[idx, key] = val
        else:
            self.budget_df = pd.concat([self.budget_df, pd.DataFrame([new_data])], ignore_index=True)
            
        if self.save_data():
            messagebox.showinfo("성공", "예산 정보가 저장되었습니다.")
            self.update_budget_view()
            # [NEW] 저장 후 현장 목록 갱신 (새로 기입한 현장명 반영)
            self.refresh_inquiry_filters()
            self.clear_budget_form()

    def _toggle_hidden_site(self, site, hide=True):
        """현장명을 드롭다운 목록에서 숨기거나 다시 표시한다."""
        if not hasattr(self, 'hidden_sites'):
            self.hidden_sites = []
        if hide:
            if site not in self.hidden_sites:
                self.hidden_sites.append(site)
            messagebox.showinfo("숨기기 완료", f"'{site}' 현장이 목록에서 숨겨졌습니다.\n다시 표시하려면 현장명 입력칸을 우클릭하세요.")
        else:
            if site in self.hidden_sites:
                self.hidden_sites.remove(site)
            messagebox.showinfo("표시 완료", f"'{site}' 현장이 목록에 다시 표시됩니다.")
        # config 저장 및 목록 갱신
        self.save_tab_config()
        self.refresh_inquiry_filters()

    def _show_all_hidden_sites(self):
        """숨긴 현장명을 모두 다시 표시한다."""
        if not getattr(self, 'hidden_sites', []):
            messagebox.showinfo("알림", "숨겨진 현장이 없습니다.")
            return
        count = len(self.hidden_sites)
        names = '\n'.join(f'  • {s}' for s in self.hidden_sites)
        confirmed = messagebox.askyesno(
            "모두 표시",
            f"현재 숨겨진 현장 {count}개:\n{names}\n\n모두 다시 표시하시겠습니까?"
        )
        if confirmed:
            self.hidden_sites.clear()
            self.save_tab_config()
            self.refresh_inquiry_filters()
            messagebox.showinfo("완료", "모든 현장이 다시 표시됩니다.")

    def delete_budget_entry(self):
        """Delete selected budget entry"""
        site = self.cb_budget_site.get().strip()
        if not site:
            messagebox.showwarning("삭제 오류", "삭제할 현장명을 입력해주세요.")
            return
            
        if messagebox.askyesno("삭제 확인", f"'{site}' 현장의 예산 정보를 삭제하시겠습니까?"):
            self.budget_df = self.budget_df[self.budget_df['Site'] != site]
            if self.save_data():
                messagebox.showinfo("성공", "삭제되었습니다.")
                self.update_budget_view()
                # [NEW] 삭제 후 현장 목록 갱신
                self.refresh_inquiry_filters()
                self.clear_budget_form()


    def clear_budget_form(self):
        """Reset budget form fields"""
        self.cb_budget_site.set('')
        self.ent_budget_revenue.delete(0, tk.END)
        if hasattr(self, 'ent_budget_unit_price'):
            self.ent_budget_unit_price.delete(0, tk.END)
        self.ent_budget_labor.delete(0, tk.END)
        self.ent_budget_material.delete(0, tk.END)
        self.ent_budget_expense.delete(0, tk.END)
        self.ent_budget_outsource.delete(0, tk.END)
        if hasattr(self, 'ent_budget_profit'):
            self.ent_budget_profit.delete(0, tk.END)
        self.ent_budget_note.delete(0, tk.END)
        # [NEW] Reset detail widgets
        if hasattr(self, 'labor_detail_widget'): self.labor_detail_widget.reset()
        if hasattr(self, 'material_detail_widget'):
            self.material_detail_widget.reset()
        if hasattr(self, 'expense_detail_widget'):
            self.expense_detail_widget.reset()
            
        # [STABILITY] Reset KPI summary
        self._update_budget_kpis()

    def on_budget_tree_select(self, event):
        """No-op - 하단 목록 제거됨"""
        pass

    def _ensure_sash_visible(self):
        """No-op as details panel is removed"""
        pass
    
    def _ensure_daily_usage_sash_visibility(self):
        """Ensure sash position is visible and properly sized using ratio"""
        try:
            if not hasattr(self, 'daily_usage_paned'):
                return
                
            # [STABILITY] Skip if we are still in the initial loading/restoration phase
            # load_tab_config will handle the initial placement.
            if not getattr(self, 'is_ready', False):
                return

            # Force multiple updates to get correct dimensions
            for _ in range(3):
                self.daily_usage_paned.update_idletasks()
            
            # Get current paned window dimensions
            total_h = self.daily_usage_paned.winfo_height()
            
            if total_h <= 0:
                print("PanedWindow height is 0, retrying...")
                self.root.after(100, self._ensure_daily_usage_sash_visibility)
                return
            
            # If locked, prioritize absolute position if valid
            if hasattr(self, 'daily_usage_sash_locked') and self.daily_usage_sash_locked:
                if hasattr(self, 'tab_config') and 'daily_usage_sash_pos' in self.tab_config:
                    target_pos = int(self.tab_config['daily_usage_sash_pos'])
                    # Ensure within bounds
                    if 50 < target_pos < total_h - 50:
                        self.daily_usage_paned.sashpos(0, target_pos)
                        print(f"Restored locked absolute position: {target_pos}")
                        return

            # Try to restore from saved ratio first
            target_ratio = 0.25  # Default ratio
            if hasattr(self, 'tab_config') and 'daily_usage_sash_ratio' in self.tab_config:
                target_ratio = self.tab_config['daily_usage_sash_ratio']
            
            # Calculate target position
            target_pos = int(total_h * target_ratio)
            
            # Apply relaxed bounds (10% to 90% of total height)
            min_pos = int(total_h * 0.1)
            max_pos = int(total_h * 0.9)
            
            if target_pos < min_pos:
                target_pos = min_pos
            elif target_pos > max_pos:
                target_pos = max_pos
            
            # Set sash position
            self.daily_usage_paned.sashpos(0, target_pos)
            
            # Verify position was set correctly
            actual_pos = self.daily_usage_paned.sashpos(0)
            actual_ratio = actual_pos / total_h if total_h > 0 else target_ratio
            
            # Update config
            if hasattr(self, 'is_ready') and self.is_ready:
                if not hasattr(self, 'tab_config'):
                    self.tab_config = {}
                self.tab_config['daily_usage_sash_ratio'] = actual_ratio
                self.tab_config['daily_usage_sash_pos'] = actual_pos
            
            # Update canvas scroll region
            self.root.after(100, self._ensure_canvas_scroll_region)
            
        except Exception as e:
            print(f"Error ensuring sash visibility: {e}")
            self.root.after(500, self._ensure_daily_usage_sash_visibility)
    
    
    
    def update_resolution_display(self):
        """Update the resolution display"""
        try:
            if hasattr(self, 'resolution_label'):
                # Get window dimensions
                window_width = self.root.winfo_width()
                window_height = self.root.winfo_height()
                
                # Get screen dimensions
                screen_width = self.root.winfo_screenwidth()
                screen_height = self.root.winfo_screenheight()
                
                # Get daily usage paned window dimensions
                if hasattr(self, 'daily_usage_paned'):
                    self.daily_usage_paned.update_idletasks()
                    pane_height = self.daily_usage_paned.winfo_height()
                    pane_width = self.daily_usage_paned.winfo_width()
                    
                    # Get current sash position and ratio
                    try:
                        sash_pos = self.daily_usage_paned.sashpos(0)
                        ratio = (sash_pos / pane_height * 100) if pane_height > 0 else 0
                        resolution_text = f"창: {window_width}x{window_height} | 화면: {screen_width}x{screen_height} | 패널: {pane_width}x{pane_height} | 경계: {ratio:.1f}%"
                    except:
                        resolution_text = f"창: {window_width}x{window_height} | 화면: {screen_width}x{screen_height} | 패널: {pane_width}x{pane_height}"
                else:
                    resolution_text = f"창: {window_width}x{window_height} | 화면: {screen_width}x{screen_height}"
                
                self.resolution_label.config(text=resolution_text)
                
                # Schedule next update
                self.root.after(500, self.update_resolution_display)
        except Exception as e:
            print(f"Error updating resolution display: {e}")
            # Still schedule next update even if there's an error
            if hasattr(self, 'resolution_label'):
                self.root.after(1000, self.update_resolution_display)
    
    def _restore_locked_position(self):
        """Restore sash to locked position"""
        try:
            if not hasattr(self, 'daily_usage_paned'): return
            
            if hasattr(self, 'tab_config'):
                total_h = self.daily_usage_paned.winfo_height()
                if total_h <= 0: return
                
                # Prioritize absolute position when locked
                if 'daily_usage_sash_pos' in self.tab_config:
                    pos = int(self.tab_config['daily_usage_sash_pos'])
                    # Only apply if it doesn't hide the bottom area completely
                    if 50 < pos < total_h - 50:
                        self.daily_usage_paned.sashpos(0, pos)
                        return
                
                # Fallback to ratio
                if 'daily_usage_sash_ratio' in self.tab_config:
                    ratio = self.tab_config['daily_usage_sash_ratio']
                    locked_pos = int(total_h * ratio)
                    self.daily_usage_paned.sashpos(0, locked_pos)
        except Exception as e:
            print(f"Error restoring locked position: {e}")
    
    def _on_daily_usage_resize(self, event):
        """Handle window resize to maintain sash ratio or absolute position if locked"""
        try:
            if not hasattr(self, 'daily_usage_paned'): return
            
            # If locked, maintain absolute position from top
            if hasattr(self, 'daily_usage_sash_locked') and self.daily_usage_sash_locked:
                self._restore_locked_position()
                return

            # Otherwise maintain ratio
            if hasattr(self, 'tab_config') and 'daily_usage_sash_ratio' in self.tab_config:
                ratio = self.tab_config['daily_usage_sash_ratio']
                total_h = self.daily_usage_paned.winfo_height()
                
                if total_h > 200:
                    new_pos = int(total_h * ratio)
                    min_pos, max_pos = 50, total_h - 50
                    new_pos = max(min_pos, min(new_pos, max_pos))
                    
                    self.daily_usage_paned.sashpos(0, new_pos)
        except Exception as e:
            print(f"Error handling resize: {e}")
            # Fallback if something went wrong during resize
            if hasattr(self, 'daily_usage_sash_locked') and self.daily_usage_sash_locked:
                self._restore_locked_position()
    
    def toggle_sash_lock(self):
        """Toggle sash lock state"""
        try:
            self.daily_usage_sash_locked = not self.daily_usage_sash_locked
            
            if self.daily_usage_sash_locked:
                # If just locked, save current state
                sash_pos = self.daily_usage_paned.sashpos(0)
                total_height = self.daily_usage_paned.winfo_height()
                
                if not hasattr(self, 'tab_config'):
                    self.tab_config = {}
                
                # Calculate and save ratio (percentage)
                ratio = sash_pos / total_height if total_height > 0 else 0.2
                self.tab_config['daily_usage_sash_ratio'] = ratio
                self.tab_config['daily_usage_sash_pos'] = sash_pos  # Keep as backup
                self.tab_config['daily_usage_sash_locked'] = True
                
                # Save configuration immediately
                self.save_tab_config(force=True)
                
                if hasattr(self, 'btn_sash_lock'):
                    self.btn_sash_lock.config(text="🔒 경계 고정됨")
                    self.btn_sash_lock.configure(style="SashLock.TButton")
                    self.style.configure("SashLock.TButton", foreground="red")
                
                print(f"Daily usage sash position LOCKED at ratio: {ratio:.3f}")
                # Start periodic monitoring
                self._start_sash_monitor()
            else:
                if not hasattr(self, 'tab_config'):
                    self.tab_config = {}
                self.tab_config['daily_usage_sash_locked'] = False
                self.save_tab_config(force=True)

                if hasattr(self, 'btn_sash_lock'):
                    self.btn_sash_lock.config(text="🔓 경계 고정")
                    self.btn_sash_lock.configure(style="TButton")
                print("Daily usage sash UNLOCKED")
                # Stop periodic monitoring
                self._stop_sash_monitor()
                
        except Exception as e:
            print(f"Error toggling sash lock: {e}")
    
    def _start_sash_monitor(self):
        """Start periodic monitoring of sash position"""
        if hasattr(self, '_sash_monitor_job'):
            self.root.after_cancel(self._sash_monitor_job)
        
        def check_sash():
            if hasattr(self, 'daily_usage_sash_locked') and self.daily_usage_sash_locked:
                self._restore_locked_position()
                # Schedule next check
                self._sash_monitor_job = self.root.after(200, check_sash)
        
        # Start monitoring
        self._sash_monitor_job = self.root.after(200, check_sash)
        print("Started sash position monitoring")
    
    def _stop_sash_monitor(self):
        """Stop periodic monitoring of sash position"""
        if hasattr(self, '_sash_monitor_job'):
            self.root.after_cancel(self._sash_monitor_job)
            delattr(self, '_sash_monitor_job')
            print("Stopped sash position monitoring")
    
    def _on_main_window_resize(self, event):
        """Handle main window resize to maintain all sash ratios"""
        try:
            # Only process for actual window resize, not widget events
            if event.widget == self.root:
                # Check if daily usage tab exists and has saved ratio
                if hasattr(self, 'daily_usage_paned') and hasattr(self, 'tab_config') and 'daily_usage_sash_ratio' in self.tab_config:
                    self.root.after(100, self._ensure_daily_usage_sash_visibility)
                
                # Check if inout tab exists and has saved ratio
                if hasattr(self, 'inout_paned') and hasattr(self, 'tab_config') and 'inout_sash_ratio' in self.tab_config:
                    self.root.after(100, self._ensure_inout_sash_visibility)
                    
        except Exception as e:
            print(f"Error handling main window resize: {e}")
    
    
    def show_error_dialog(self, title, message):
        """Show a custom error dialog with draggable text"""
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("600x400")
        dialog.resizable(True, True)
        
        # Make dialog modal
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (600 // 2)
        y = (dialog.winfo_screenheight() // 2) - (400 // 2)
        dialog.geometry(f"600x400+{x}+{y}")
        
        # Main frame
        main_frame = ttk.Frame(dialog)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Title label
        title_label = ttk.Label(main_frame, text=title, font=('Malgun Gothic', 12, 'bold'))
        title_label.pack(pady=(0, 10))
        
        # Text widget with scrollbar for error message
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill='both', expand=True)
        
        text_widget = tk.Text(text_frame, wrap='word', font=('Malgun Gothic', 10))
        scrollbar = ttk.Scrollbar(text_frame, orient='vertical', command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Insert error message
        text_widget.insert('1.0', message)
        text_widget.configure(state='normal')  # Allow selection and copying
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill='x', pady=(10, 0))
        
        # Copy button
        def copy_text():
            dialog.clipboard_clear()
            selected_text = text_widget.get('sel.first', 'sel.last') if text_widget.tag_ranges('sel') else text_widget.get('1.0', 'end')
            dialog.clipboard_append(selected_text)
        
        copy_btn = ttk.Button(button_frame, text="복사하기", command=copy_text)
        copy_btn.pack(side='left', padx=5)
        
        # Close button
        close_btn = ttk.Button(button_frame, text="닫기", command=dialog.destroy)
        close_btn.pack(side='right', padx=5)
        
        # Focus on text widget
        text_widget.focus_set()
        
        # Bind Escape key to close
        dialog.bind('<Escape>', lambda e: dialog.destroy())
        
        # Wait for dialog to close
        dialog.wait_window()

    def on_daily_usage_tree_select(self, event):
        """No-op: details panel removed"""
        pass

    def delete_selected_site(self):
        """Remove the selected site from the suggestion list"""
        site = self.cb_daily_site.get().strip()
        if not site:
            messagebox.showwarning("선택 오류", "삭제할 현장명을 선택해주세요.")
            return
            
        if site in self.sites:
            if messagebox.askyesno("삭제 확인", f"'{site}' 현장명을 목록에서 삭제하시겠습니까?\n(기존 기록은 삭제되지 않습니다.)"):
                self.sites.remove(site)
                self.cb_daily_site['values'] = self.sites
                self.cb_daily_site.set('')
                self.save_tab_config()
                messagebox.showinfo("완료", "현장명이 삭제되었습니다.")
        else:
            messagebox.showinfo("알림", "목록에 없는 현장명입니다.")

    def calculate_rtk_total(self):
        """Calculate total RTK usage"""
        try:
            total = 0
            rtk_categories = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타"]
            for category in rtk_categories:
                value = self.rtk_entries[category].get()
                if value:
                    total += float(value)
            self.rtk_entries["총계"].config(state='normal')
            self.rtk_entries["총계"].delete(0, tk.END)
            # Format as integer if possible, else float
            if total == int(total):
                self.rtk_entries["총계"].insert(0, str(int(total)))
            else:
                self.rtk_entries["총계"].insert(0, f"{total:.1f}")
            self.rtk_entries["총계"].config(state='readonly')
        except ValueError:
            pass  # Ignore invalid input during typing
    
    def sync_film_with_quantity(self, event=None):
        """Sync film count with quantity if method is RT"""
        try:
            method = self.cb_daily_test_method.get().strip().upper()
            if "RT" == method or method.startswith("RT "):
                qty = self.ent_daily_test_amount.get().strip()
                if hasattr(self, 'ent_film_count'):
                    self.ent_film_count.delete(0, tk.END)
                    self.ent_film_count.insert(0, qty)
        except Exception as e:
            print(f"Error syncing film with quantity: {e}")

    def format_entry_with_commas(self, event, entry):
        """Autoformat numbers with commas as the user types or leaves the field"""
        try:
            val = entry.get().strip().replace(',', '')
            if not val: return
            
            # Use float or int depending on content
            if '.' in val:
                num = float(val)
                formatted = f"{num:,.1f}"
                if formatted.endswith('.0'): formatted = formatted[:-2]
            else:
                num = int(val)
                formatted = f"{num:,}"
                
            entry.delete(0, tk.END)
            entry.insert(0, formatted)
        except: pass

    def update_daily_test_fee_calc(self, event=None):
        """Auto-calculate Inspection Fee = (Amount * Unit Price) + Travel Expense + Meal Cost"""
        try:
            def get_f(entry):
                try:
                    val = entry.get().strip().replace(',', '')
                    return float(val) if val else 0.0
                except: return 0.0

            amount = get_f(self.ent_daily_test_amount)
            price = get_f(self.ent_daily_unit_price)
            travel = get_f(self.ent_daily_travel_cost)
            meal = get_f(self.ent_daily_meal_cost)
            
            calc_fee = (amount * price) + travel + meal
            
            # Update the fee field with commas
            self.ent_daily_test_fee.delete(0, tk.END)
            self.ent_daily_test_fee.insert(0, f"{calc_fee:,.0f}")
        except:
            pass

    def sync_worker_times(self):
        """작업자 1의 설정(주야/작업시간/OT)을 모든 작업자와 동기화"""
        try:
            # Get values from Worker 1
            if not hasattr(self, 'worker_group1'):
                return
                
            master_group = self.worker_group1
            wt1 = master_group.ent_worktime.get().strip()
            ot1 = master_group.ent_ot.get().strip()
            shift1 = master_group.cb_shift.get()
            
            if not wt1 and not ot1:
                messagebox.showwarning("입력 필요", "작업자 1의 작업시간이나 OT를 입력해주세요.")
                return

            for i in range(2, 11):
                group_attr = f'worker_group{i}'
                
                # Check if this worker slot exists
                if not hasattr(self, group_attr):
                    continue
                    
                target_group = getattr(self, group_attr)
                
                # Only apply to workers with names selected
                target_name = target_group.get_worker().strip()
                if not target_name:
                    continue

                # Sync Shift
                target_group.cb_shift.set(shift1)

                # Sync Work Time
                target_group.ent_worktime.set(wt1)
                
                # Sync OT
                target_group.set_ot(ot1)
            
            messagebox.showinfo("완료", "작업자 1의 설정(주야/시간/OT)이 성명이 입력된 모든 작업자에게 적용되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"동기화 중 오류가 발생했습니다: {e}")

    def add_daily_usage_entry(self):
        """Add a daily usage entry"""
        try:
            date_str = self.ent_daily_date.get()
            site = self.cb_daily_site.get().strip()
            mat_name = self.cb_daily_material.get()
            film_count_str = self.ent_film_count.get()
            note = "" # Note widget removed per user request
            ndt_usage = {} # Initialize NDT usage dictionary
            
            # Validate film count
            try:
                film_count = float(film_count_str) if film_count_str else 0.0
            except ValueError:
                messagebox.showwarning("입력 오류", "필름매수는 숫자여야 합니다.")
                return
            
            # Get RTK usage values
            rtk_values = {}
            total_usage = 0
            rtk_categories = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타"]
            
            for category in rtk_categories:
                value_str = self.rtk_entries[category].get()
                try:
                    value = float(value_str) if value_str else 0.0
                    rtk_values[category] = value
                    total_usage += value
                except ValueError:
                    messagebox.showwarning("입력 오류", f"{category} 사용량은 숫자여야 합니다.")
                    return
            
            # Validation
            if not date_str:
                messagebox.showwarning("입력 오류", "날짜를 입력해주세요.")
                return
            
            if not site:
                messagebox.showwarning("입력 오류", "현장을 입력해주세요.")
                return
            
            # Robust MaterialID Lookup (품목명 없으면 건너뜀)
            mat_id = None
            pure_mat_name = ""

            if mat_name:
                for _, m_row in self.materials_df.iterrows():
                    m_name = str(m_row.get('품목명', ''))
                    m_model = str(m_row.get('모델명', '')).replace('nan', '').replace('None', '')
                    m_sn = str(m_row.get('SN', '')).replace('nan', '').replace('None', '')
                    
                    m_display = m_name
                    if m_model: m_display += f" - {m_model}"
                    if m_sn: m_display += f" (SN: {m_sn})"
                    
                    if m_display == mat_name:
                        mat_id = m_row['MaterialID']
                        pure_mat_name = m_name
                        break
                
                # Fallback for manual entry if exact match not found
                if mat_id is None:
                    pure_mat_name = mat_name
                    if " - " in pure_mat_name:
                        pure_mat_name = pure_mat_name.split(" - ")[0]
                    if " (SN: " in pure_mat_name:
                        pure_mat_name = pure_mat_name.split(" (SN: ")[0]
                    
                    # Try finding by name only
                    mat_rows = self.materials_df[self.materials_df['품목명'] == pure_mat_name]
                    if not mat_rows.empty:
                        mat_id = mat_rows['MaterialID'].values[0]
            
            # Parse date and combine with current time for transaction accuracy
            try:
                selected_date = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                # Use current time to provide full timestamp for In/Out records
                usage_datetime = datetime.datetime.combine(selected_date.date(), datetime.datetime.now().time())
            except ValueError:
                messagebox.showwarning("입력 오류", "날짜 형식이 올바르지 않습니다. (YYYY-MM-DD)")
                return
            
            if mat_name and mat_id is None:
                # Material doesn't exist in database, create it
                new_mat_id = self.materials_df['MaterialID'].max() + 1 if not self.materials_df.empty else 1
                
                # More inclusive brand-to-category mapping for new materials
                default_cat = 'OTHER'
                name_up = pure_mat_name.upper()
                if any(k in name_up for k in ['CARESTREAM', 'AGFA', 'FUJI', 'KODAK', 'STRUCTURIX', 'FILM', '필름']):
                    default_cat = 'FILM'
                elif any(k in name_up for k in ['침투제', '세척제', '현상제', '자분', '페인트', 'NABAKEM', 'MAGNAFLUX']):
                    default_cat = 'NDT_CHEM'
                
                new_material = {
                    'MaterialID': new_mat_id,
                    '품목명': pure_mat_name,
                    '관리품번': '',
                    '품목군코드': default_cat,
                    '규격': '',
                    '관리단위': 'EA',
                    '재고위치': '',
                    '공급업체': '',
                    '제조사': '',
                    '재고하한': 10,
                    '수량': 0
                }
                
                self.materials_df = pd.concat([self.materials_df, pd.DataFrame([new_material])], ignore_index=True)
                self.save_data()
                mat_id = new_mat_id
            # else: mat_name 없거나 mat_id가 이미 설정됨 — 그대로 유지
            
            # 1. Collect all worker data from UI
            raw_worker_data = []
            
            # Helper to get worker info from UI using the new group methods
            for i in range(1, 11):
                group_attr = f'worker_group{i}'
                if hasattr(self, group_attr):
                    group = getattr(self, group_attr)
                    u_name = group.get_worker().strip()
                    u_time = group.get_time().strip() # This now includes (Shift) prefix
                    u_ot = group.get_ot().strip()
                    if u_name:
                        raw_worker_data.append({'name': u_name, 'time': u_time, 'ot': u_ot})
            
            # 2. Sort workers alphabetically by name (가나다순)
            sorted_workers = sorted(raw_worker_data, key=lambda x: x['name'])
            
            # 3. Prepare final variables for User1-10 (fill up to 10 slots)
            # This ensures we always have 10 elements to unpack/reference
            final_worker_list = sorted_workers + [{'name': '', 'time': '', 'ot': ''}] * (10 - len(sorted_workers))
            
            manager_val = final_worker_list[0]['name']
            wt1_val     = final_worker_list[0]['time']
            ot1_val     = final_worker_list[0]['ot']
            
            user2_val   = final_worker_list[1]['name']
            wt2_val     = final_worker_list[1]['time']
            ot2_val     = final_worker_list[1]['ot']
            
            user3_val   = final_worker_list[2]['name']
            wt3_val     = final_worker_list[2]['time']
            ot3_val     = final_worker_list[2]['ot']
            
            user4_val   = final_worker_list[3]['name']
            wt4_val     = final_worker_list[3]['time']
            ot4_val     = final_worker_list[3]['ot']
            
            user5_val   = final_worker_list[4]['name']
            wt5_val     = final_worker_list[4]['time']
            ot5_val     = final_worker_list[4]['ot']
            
            user6_val   = final_worker_list[5]['name']
            wt6_val     = final_worker_list[5]['time']
            ot6_val     = final_worker_list[5]['ot']
            
            user7_val   = final_worker_list[6]['name']
            wt7_val     = final_worker_list[6]['time']
            ot7_val     = final_worker_list[6]['ot']
            
            user8_val   = final_worker_list[7]['name']
            wt8_val     = final_worker_list[7]['time']
            ot8_val     = final_worker_list[7]['ot']
            
            user9_val   = final_worker_list[8]['name']
            wt9_val     = final_worker_list[8]['time']
            ot9_val     = final_worker_list[8]['ot']
            
            user10_val  = final_worker_list[9]['name']
            wt10_val    = final_worker_list[9]['time']
            ot10_val    = final_worker_list[9]['ot']
            
            all_workers = ", ".join([w['name'] for w in sorted_workers])

            equip_name = self.cb_daily_equip.get().strip() if hasattr(self, 'cb_daily_equip') else ''
            test_method = self.cb_daily_test_method.get().strip() if hasattr(self, 'cb_daily_test_method') else ''
            
            try:
                def get_f(ent):
                    try:
                        val = ent.get().strip().replace(',', '')
                        return float(val) if val else 0.0
                    except: return 0.0
                
                test_amount = get_f(self.ent_daily_test_amount)
                unit_price = get_f(self.ent_daily_unit_price)
                travel_cost = get_f(self.ent_daily_travel_cost)
                meal_cost = get_f(self.ent_daily_meal_cost)
                test_fee = get_f(self.ent_daily_test_fee)
            except:
                test_amount = unit_price = travel_cost = meal_cost = test_fee = 0.0

            # --- Collect Vehicle Inspection Data ---
            v_infos = []
            v_mileages = []
            v_checks = []
            v_remarks = []
            
            for v_key, v_widget in self.vehicle_inspections.items():
                v_data = v_widget.get_data()
                v_infos.append(v_data.get('vehicle_info', ''))
                v_mileages.append(v_data.get('mileage', ''))
                v_remarks.append(v_data.get('remarks', ''))
                
                # Format checks: Fuel:V, Clean:X etc.
                checks = []
                for label, key in [("주유", "Fuel"), ("청결", "Clean"), ("엔진오일", "Oil"), ("타이어", "Tire"), ("전조등", "Light")]:
                    if v_data.get(key):
                        checks.append(label)
                
                # Always append to maintain alignment with other fields
                v_checks.append(", ".join(checks))
            
            vehicle_info_str = ", ".join([v for v in v_infos if v])
            mileage_str = ", ".join([v for v in v_mileages if v])
            inspection_str = "; ".join([v for v in v_checks if v])
            v_remarks_str = ", ".join([v for v in v_remarks if v])

            new_entry = {
                'Date': selected_date,
                'Site': site.strip(),
                'MaterialID': mat_id,
                '장비명': equip_name,
                '검사방법': test_method,
                '검사량': test_amount,
                '단가': unit_price,
                '출장비': travel_cost,
                '일식': meal_cost,
                '검사비': test_fee,
                'FilmCount': film_count,
                'Usage': total_usage,
                'Note': note,
                'EntryTime': datetime.datetime.now(),
                'User': manager_val,
                'WorkTime': wt1_val,
                'OT': ot1_val,
                'User2': user2_val,
                'WorkTime2': wt2_val,
                'OT2': ot2_val,
                'User3': user3_val,
                'WorkTime3': wt3_val,
                'OT3': ot3_val,
                'User4': user4_val,
                'WorkTime4': wt4_val,
                'OT4': ot4_val,
                'User5': user5_val,
                'WorkTime5': wt5_val,
                'OT5': ot5_val,
                'User6': user6_val,
                'WorkTime6': wt6_val,
                'OT6': ot6_val,
                'User7': user7_val,
                'WorkTime7': wt7_val,
                'OT7': ot7_val,
                'User8': user8_val,
                'WorkTime8': wt8_val,
                'OT8': ot8_val,
                'User9': user9_val,
                'WorkTime9': wt9_val,
                'OT9': ot9_val,
                'User10': user10_val,
                'WorkTime10': wt10_val,
                'OT10': ot10_val,
                '차량번호': vehicle_info_str,
                '주행거리': mileage_str,
                '차량점검': inspection_str,
                '차량비고': v_remarks_str
            }
            
            # Add RTK values to the entry
            for category, value in rtk_values.items():
                new_entry[f'RTK_{category}'] = value
            
            # Add NDT materials values to the entry
            ndt_materials = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
            
            for i, material in enumerate(ndt_materials):
                value_str = self.ndt_entries[material].get()
                try:
                    value = float(value_str) if value_str else 0.0
                    # Remove all whitespace from material name for the key
                    m_key = "".join(material.split())
                    new_entry[f'NDT_{m_key}'] = value
                    if value > 0:
                        ndt_usage[material] = value
                except ValueError:
                    m_key = "".join(material.split())
                    new_entry[f'NDT_{m_key}'] = 0.0
            
            # Final key cleanup for the entire entry to be 100% safe
            # Ensure all keys are strings and values are properly formatted
            final_entry = {}
            for k, v in new_entry.items():
                clean_key = str(k).strip()
                if isinstance(v, (int, float)):
                    final_entry[clean_key] = float(v)
                else:
                    final_entry[clean_key] = str(v) if v is not None else ""
            
            # Stock deduction via transaction
            stock_info = ""
            
            # Keyword mapping for NDT materials (PT/MT Chemicals)
            ndt_keywords = {
                "형광자분": ["형광자분", "FLUOR", "MAGNETIC", "GLO"],
                "흑색자분": ["흑색", "자분", "BLACK", "MAGNETIC", "7HF"],
                "백색페인트": ["백색", "페인트", "WHITE", "PAINT", "CONTRAST", "WCP"],
                "침투제": ["침투제", "PENETRANT", "VP", "SKL", "NP"],
                "세척제": ["세척제", "CLEANER", "SKC", "CLEAN", "NC"],
                "현상제": ["현상제", "DEVELOPER", "SKD", "DEV", "ND"],
                "형광침투제": ["형광", "FLUOR", "GLO"]
            }
            
            # 1. Deduct main selection (FILM or Item based on Inspection Amount)
            # UNIFIED LOGIC: Total Deduction = Film Count + RTK Total (total_usage)
            # Inspection Quantity (test_amount) is EXCLUDED from deduction as per user request ("검사량 빼줘").
            total_deduction = film_count + total_usage
            if mat_id is not None and total_deduction > 0:
                is_film = False
                mat_row = self.materials_df[self.materials_df['MaterialID'] == mat_id]
                category = ""
                if not mat_row.empty:
                    category = str(mat_row.iloc[0].get('품목군코드', '')).upper()
                
                mat_name_upper = pure_mat_name.upper()
                
                # Very broad FILM detection
                if any(k in category for k in ['FILM', '필름', 'FLM', 'RT']):
                    is_film = True
                elif any(k in mat_name_upper for k in ['CARESTREAM', 'AGFA', 'FUJI', 'KODAK', 'STRUCTURIX', 'FILM', '필름', 'IX100', 'IX80', 'IX50', 'FOMA']):
                    is_film = True
                elif test_method == "RT":
                    is_film = True
                
                # Create transaction for total deduction
                new_transaction = {
                    'Date': usage_datetime,
                    'MaterialID': mat_id,
                    'Type': 'OUT',
                    'Quantity': -total_deduction,
                    'User': all_workers,
                    'Site': site,
                    'Note': f'{site} 현장 사용 (자동 차감)',
                    '차량번호': final_entry.get('차량번호', ''),
                    '주행거리': final_entry.get('주행거리', ''),
                    '차량점검': final_entry.get('차량점검', ''),
                    '차량비고': final_entry.get('차량비고', '')
                }
                self.transactions_df = pd.concat([self.transactions_df, pd.DataFrame([new_transaction])], ignore_index=True)
                # Ensure columns stay consistent
                self.transactions_df.columns = [str(c).strip() for c in self.transactions_df.columns]
                
                # Build summary message
                parts = []
                if film_count > 0: parts.append(f"{film_count:.1f}(필름)")
                # if test_amount > 0: parts.append(f"{test_amount:.1f}(검사량)") # Excluded from deduction
                if total_usage > 0: parts.append(f"{total_usage:.1f}(재촬영)")
                summary = "+".join(parts)
                
                stock_info += f"\n• {pure_mat_name}: {summary} = {total_deduction:.1f} 차감"
            else:
                 # No deduction if total is 0
                 pass

            # 2. Deduct NDT materials (PT/MT Chemicals)
            for input_name, qty in ndt_usage.items():
                # Fallback mapping for renamed items
                search_names = [input_name]
                if input_name == "흑색자분": search_names.append("자분")
                elif input_name == "백색페인트": search_names.append("페인트")
                elif input_name == "형광침투제": search_names.append("형광")
                
                ndt_mat_rows = pd.DataFrame()
                
                # Step 1: Try exact match for any of the search names (품목명 or 모델명)
                for s_name in search_names:
                    matches = self.materials_df[
                        (self.materials_df['품목명'] == s_name) | 
                        (self.materials_df['모델명'] == s_name)
                    ]
                    if not matches.empty:
                        ndt_mat_rows = matches
                        break
                
                # Step 2: Try finding using keywords from ndt_keywords mapping
                if ndt_mat_rows.empty:
                    keywords = ndt_keywords.get(input_name, [input_name])
                    for kw in keywords:
                        matches = self.materials_df[
                            self.materials_df['품목명'].astype(str).str.contains(kw, na=False, case=False, regex=False) |
                            self.materials_df['모델명'].astype(str).str.contains(kw, na=False, case=False, regex=False)
                        ]
                        if not matches.empty:
                            ndt_mat_rows = matches
                            break
                
                # Step 3: Fallback to partial match of original name
                if ndt_mat_rows.empty:
                    for s_name in search_names:
                        matches = self.materials_df[
                            self.materials_df['품목명'].astype(str).str.contains(s_name, na=False, case=False, regex=False) |
                            self.materials_df['모델명'].astype(str).str.contains(s_name, na=False, case=False, regex=False)
                        ]
                        if not matches.empty:
                            ndt_mat_rows = matches
                            break
                    
                if not ndt_mat_rows.empty:
                    # Check category for PT/MT chemicals
                    ndt_mat_row = ndt_mat_rows.iloc[0]
                    ndt_mat_id = ndt_mat_row['MaterialID']
                    actual_mat_name = ndt_mat_row['품목명']
                    category = str(ndt_mat_row.get('품목군코드', '')).upper()
                    
                    # Check if it belongs to PT or MT chemicals
                    is_chemical = False
                    category_upper = category.upper()
                    name_upper = str(actual_mat_name).upper()
                    
                    if any(k in category_upper for k in ['PT', 'MT', '약품', '소모품', 'CHEM', 'NDT']):
                        is_chemical = True
                    # Fallback: specific names if category is vague
                    elif any(k in name_upper for k in ['침투제', '세척제', '현상제', '자분', '페인트', '흑색자분', '백색페인트', '형광', 'DEVELOPER', 'CLEANER', 'PENETRANT', 'MT', 'PT', 'NABAKEM', 'MAGNAFLUX']):
                        is_chemical = True
                    
                    if is_chemical:
                        ndt_transaction = {
                            'Date': usage_datetime,
                            'MaterialID': ndt_mat_id,
                            'Type': 'OUT',
                            'Quantity': -qty,
                            'User': all_workers,
                            'Site': site,
                            'Note': f'{site} 현장 사용 (자동 차감)'
                        }
                        self.transactions_df = pd.concat([self.transactions_df, pd.DataFrame([ndt_transaction])], ignore_index=True)
                        stock_info += f"\n• {actual_mat_name}: {qty:.1f} 차감 (-)"
                    else:
                        stock_info += f"\n• {actual_mat_name}: 재고 차감 대상 아님 (구분:'{category}', 약품 아님)"
                else:
                    # Auto-register missing NDT material
                    new_ndt_id = self.materials_df['MaterialID'].max() + 1 if not self.materials_df.empty else 1
                    
                    new_ndt_material = {
                        'MaterialID': new_ndt_id,
                        '품목명': input_name,
                        '관리품번': '',
                        '품목군코드': 'NDT_CHEM',
                        '규격': '',
                        '관리단위': 'EA',
                        '재고위치': '',
                        '공급업체': '',
                        '제조사': '',
                        '재고하한': 5,
                        '수량': 0
                    }
                    
                    self.materials_df = pd.concat([self.materials_df, pd.DataFrame([new_ndt_material])], ignore_index=True)
                    self.save_data()
                    
                    ndt_transaction = {
                        'Date': usage_datetime,
                        'MaterialID': new_ndt_id,
                        'Type': 'OUT',
                        'Quantity': -qty,
                        'User': all_workers,
                        'Site': site,
                        'Note': f'{site} 현장 사용 (자동 등록 및 차감)'
                    }
                    self.transactions_df = pd.concat([self.transactions_df, pd.DataFrame([ndt_transaction])], ignore_index=True)
                    stock_info += f"\n• {input_name}: 자동 등록 및 {qty:.1f} 차감 (-)"
            
            self.daily_usage_df = pd.concat([self.daily_usage_df, pd.DataFrame([final_entry])], ignore_index=True)
            # Maintain clean headers
            self.daily_usage_df.columns = [str(c).strip() for c in self.daily_usage_df.columns]
            
            # Ensure Date column is consistently formatted
            if 'Date' in self.daily_usage_df.columns:
                self.daily_usage_df['Date'] = pd.to_datetime(self.daily_usage_df['Date'], errors='coerce').dt.strftime('%Y-%m-%d')
            
            # Count transactions effectively added in this call
            # (This is a bit hard with the current structure, but we can check if self.transactions_df length increased)
            # Actually, let's just make sure we save and refresh.
            
            if self.save_data():
                # Set Daily Usage filters to the entry date to show relevant columns only
                if hasattr(self, 'ent_daily_start_date'): 
                    try:
                        self.ent_daily_start_date.set_date(selected_date)
                    except:
                        self.ent_daily_start_date.delete(0, tk.END)
                        self.ent_daily_start_date.insert(0, date_str)
                if hasattr(self, 'ent_daily_end_date'): 
                    try:
                        self.ent_daily_end_date.set_date(selected_date)
                    except:
                        self.ent_daily_end_date.delete(0, tk.END)
                        self.ent_daily_end_date.insert(0, date_str)
                if hasattr(self, 'cb_daily_filter_site'): self.cb_daily_filter_site.set("전체")
                if hasattr(self, 'cb_daily_filter_material'): self.cb_daily_filter_material.set("전체")
                
                self.update_daily_usage_view()
                
                # Reset filters in In/Out tab to ensure the new transaction is visible
                if hasattr(self, 'cb_trans_filter_mat'):
                    self.cb_trans_filter_mat.set("전체")
                if hasattr(self, 'cb_trans_filter_site'):
                    self.cb_trans_filter_site.set("전체")
                    
                self.update_stock_view() # Refresh stock tab to reflect deductions
                self.update_transaction_view() # Ensure In/Out tab is refreshed
                
                # Update site list if it's a new site
                if site not in self.sites:
                    self.sites.append(site)
                    self.sites.sort()
                    self.cb_daily_site['values'] = self.sites
                    self.save_tab_config()
                    stock_info += f"\n• 신규 현장 '{site}'이 목록에 저장되었습니다."
                
                # Update user list if any new users entered
                for i in range(1, 11):
                    cb_attr = f'cb_daily_user{i}' if i > 1 else 'cb_daily_user'
                    if hasattr(self, cb_attr):
                        u_name = getattr(self, cb_attr).get().strip()
                        # Extract actual name from "(Shift) Name" format
                        import re
                        actual_name = u_name
                        match = re.match(r"\((주간|야간|휴일)\)\s*(.*)", u_name)
                        if match:
                            # Extract only the name part (group 2)
                            actual_name = match.group(2).strip()
                        
                        # Only add if there's a real name (not empty or just shift)
                        if actual_name and actual_name not in self.users:
                            self.users.append(actual_name)
                            self.users.sort()
                            self.refresh_ui_for_list_change('users')
                            self.save_tab_config()
                            stock_info += f"\n• 신규 담당자 '{actual_name}'이 목록에 저장되었습니다."
                
                # Update stock view and material lists
                self.update_stock_view()
                self.update_material_combo()
                self.update_registration_combos()
                
                final_msg = f"{site}의 {pure_mat_name} 사용 기록이 저장되었습니다."
                if stock_info:
                    final_msg += f"\n\n[재고 차감 내역]:{stock_info}"
                else:
                    final_msg += "\n\n(재고 차감 대상이 없어 입출고 내역이 생성되지 않았습니다.)"
                    
                messagebox.showinfo("완료", final_msg)
                self.refresh_inquiry_filters()
                self.update_daily_usage_view()
                # 월별 탭 필터를 저장된 날짜에 맞추고 갱신
                try:
                    if hasattr(self, 'cb_filter_year'):
                        self.cb_filter_year.set(str(selected_date.year))
                    if hasattr(self, 'cb_filter_month'):
                        self.cb_filter_month.set(str(selected_date.month))
                except Exception:
                    pass
                self.update_monthly_usage_view() # Also refresh monthly summary
                
                # Report view was removed/renamed, safe to skip or call appropriate method
                if hasattr(self, 'update_report_view'):
                    self.update_report_view()
                elif hasattr(self, 'update_budget_view'):
                    # Could potentially refresh budget if needed, but not strictly required here
                    pass
                
                # Switch to In/Out Management tab (index 1) to see the transactions
                try:
                    self.notebook.select(1)
                except:
                    pass
                    
                # Clear fields (requested by user to always keep empty except date)
                self.cb_daily_site.set('') 
                self.cb_daily_equip.set('')
                self.cb_daily_material.set('')
                self.cb_daily_test_method.set('')
                
                self.ent_daily_test_amount.delete(0, tk.END)
                self.ent_daily_unit_price.delete(0, tk.END)
                self.ent_daily_travel_cost.delete(0, tk.END)
                self.ent_daily_meal_cost.delete(0, tk.END)
                self.ent_daily_test_fee.delete(0, tk.END)
                self.ent_film_count.delete(0, tk.END)
                
                for i in range(1, 11):
                    cb_attr = f'cb_daily_user{i}' if i > 1 else 'cb_daily_user'
                    if hasattr(self, cb_attr):
                        getattr(self, cb_attr).set('') # Clear user combobox
                    
                    wt_attr = f'ent_worktime{i}' if i > 1 else 'ent_worktime1'
                    ot_attr = f'ent_ot{i}' if i > 1 else 'ent_ot1'
                    if hasattr(self, wt_attr): getattr(self, wt_attr).delete(0, tk.END)
                    if hasattr(self, ot_attr): getattr(self, ot_attr).delete(0, tk.END)

                for category in rtk_categories:
                    self.rtk_entries[category].delete(0, tk.END)
                for material in ndt_materials:
                    self.ndt_entries[material].delete(0, tk.END)
                
                self.rtk_entries["총계"].config(state='normal')
                self.rtk_entries["총계"].delete(0, tk.END)
                self.rtk_entries["총계"].config(state='readonly')
            else:
                # Save returned False - do not clear fields
                pass # The error message is already shown by save_data

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            self.show_error_dialog("저장 오류", f"기록 저장 중 오류가 발생했습니다:\n{e}\n\n상세 정보:\n{error_details}")
    
    def refresh_inquiry_filters(self):
        """Unified method to populate inquiry filter dropdowns and sync with autocomplete lists"""
        import re
        marker_pattern = MARKER_PATTERN
        
        # 1. Collect and Merge Site Info
        raw_sites = set()
        if not self.daily_usage_df.empty and 'Site' in self.daily_usage_df.columns:
            raw_sites.update(self.daily_usage_df['Site'].dropna().astype(str).str.strip().tolist())
        if hasattr(self, 'sites'):
            raw_sites.update(self.sites)
        # [NEW] budget_df에 저장된 현장명도 포함
        if hasattr(self, 'budget_df') and not self.budget_df.empty and 'Site' in self.budget_df.columns:
            raw_sites.update(self.budget_df['Site'].dropna().astype(str).str.strip().tolist())
        
        unique_sites = sorted([s for s in raw_sites 
                                if s and str(s).lower() != 'nan'
                                and s not in getattr(self, 'hidden_sites', [])])  # [FIX] 숨긴 현장 제외
        self.sites[:] = unique_sites # Update master list in-place
        
        if hasattr(self, 'cb_daily_filter_site'):
            self.cb_daily_filter_site['values'] = ['전체'] + unique_sites
            if not self.cb_daily_filter_site.get(): self.cb_daily_filter_site.set('전체')
        if hasattr(self, 'cb_filter_site_monthly'):
            self.cb_filter_site_monthly['values'] = ['전체'] + unique_sites
        # [FIX] 공사실행예산서 상단 현장 콤보박스 → budget_df 현장만 표시
        # (삭제 후 목록에서 즉시 사라지게 하기 위해 daily_usage_df 현장은 제외)
        if hasattr(self, 'cb_budget_site'):
            budget_sites = []
            if hasattr(self, 'budget_df') and not self.budget_df.empty and 'Site' in self.budget_df.columns:
                budget_sites = sorted([
                    s for s in self.budget_df['Site'].dropna().astype(str).str.strip().unique()
                    if s and str(s).lower() != 'nan'
                ])
            self.budget_sites[:] = budget_sites  # [FIX] 자동완성 목록 동기화
            current_val = self.cb_budget_site.get()
            self.cb_budget_site['values'] = budget_sites
            if current_val: self.cb_budget_site.set(current_val)
        # 하단 실적 조회 콤보박스 → 모든 현장(daily_usage_df + budget_df) 표시
        if hasattr(self, 'cb_budget_view_site'):
            self.cb_budget_view_site['values'] = unique_sites

        # 2. Collect and Merge Material Info
        raw_materials = set()
        if not self.daily_usage_df.empty and 'MaterialID' in self.daily_usage_df.columns:
            unique_mat_ids = self.daily_usage_df['MaterialID'].dropna().unique()
            for mat_id in unique_mat_ids:
                name = self.get_material_display_name(mat_id)
                if name: raw_materials.add(name)
        if hasattr(self, 'materials_display_list'):
            raw_materials.update(self.materials_display_list)
        
        unique_materials = sorted([m for m in raw_materials if m])
        self.materials_display_list[:] = unique_materials # Update master list in-place
        
        if hasattr(self, 'cb_daily_filter_material'):
            self.cb_daily_filter_material['values'] = ['전체'] + unique_materials
            if not self.cb_daily_filter_material.get(): self.cb_daily_filter_material.set('전체')
        if hasattr(self, 'cb_filter_material_monthly'):
            self.cb_filter_material_monthly['values'] = ['전체'] + unique_materials

        # 3. Equipment dropdown
        if hasattr(self, 'cb_daily_filter_equipment'):
            raw_equip = []
            if not self.daily_usage_df.empty and '장비명' in self.daily_usage_df.columns:
                raw_equip = self.daily_usage_df['장비명'].dropna().astype(str).str.strip().unique().tolist()
            
            # Combine with catalog if needed
            unique_equip = sorted(list(set([e for e in raw_equip if e and str(e).lower() != 'nan'])))
            self.cb_daily_filter_equipment['values'] = ['전체'] + unique_equip
            if not self.cb_daily_filter_equipment.get(): self.cb_daily_filter_equipment.set('전체')

        # 4. Worker dropdown
        if hasattr(self, 'cb_daily_filter_worker'):
            worker_cols = ['User', 'User2', 'User3', 'User4', 'User5', 'User6', 'User7', 'User8', 'User9', 'User10']
            all_workers_raw = set()
            if not self.daily_usage_df.empty:
                for col in worker_cols:
                    if col in self.daily_usage_df.columns:
                        all_workers_raw.update(self.daily_usage_df[col].dropna().astype(str).unique().tolist())
            
            # Include static list
            if hasattr(self, 'users'): all_workers_raw.update(self.users)
            
            clean_workers_set = set()
            for w in all_workers_raw:
                w_str = str(w).strip()
                if not w_str or w_str.lower() in ['nan', '0.0', 'none']: continue
                cleaned = marker_pattern.sub('', w_str).strip()
                cleaned = " ".join(cleaned.split())
                if cleaned: clean_workers_set.add(cleaned)

            unique_workers = sorted(list(clean_workers_set))
            self.users[:] = unique_workers # Update master list in-place
            self.cb_daily_filter_worker['values'] = ['전체'] + unique_workers
            if not self.cb_daily_filter_worker.get(): self.cb_daily_filter_worker.set('전체')

        # 5. Vehicle dropdown
        raw_vehicles = set()
        if not self.daily_usage_df.empty and '차량번호' in self.daily_usage_df.columns:
            v_list = self.daily_usage_df['차량번호'].dropna().astype(str).str.strip().tolist()
            for v in v_list:
                cleaned = self.clean_nan(v)
                if cleaned: raw_vehicles.add(cleaned)
        if hasattr(self, 'vehicles'):
            raw_vehicles.update(self.vehicles)
            
        unique_vehicles = sorted(list(raw_vehicles))
        self.vehicles[:] = unique_vehicles # Sync with master list
        
        if hasattr(self, 'cb_daily_filter_vehicle'):
            self.cb_daily_filter_vehicle['values'] = ['전체'] + unique_vehicles
            if not self.cb_daily_filter_vehicle.get(): self.cb_daily_filter_vehicle.set('전체')
        if hasattr(self, 'cb_sales_filter_site'):
            self.cb_sales_filter_site['values'] = ['전체'] + self.sites
            if not self.cb_sales_filter_site.get(): self.cb_sales_filter_site.set('전체')
        if hasattr(self, 'cb_trans_filter_vehicle'):
            self.cb_trans_filter_vehicle['values'] = ['전체'] + unique_vehicles

    def update_daily_usage_view(self):
        """Update the daily usage treeview with smarter filters and shift classification"""
        import re
        marker_pattern = MARKER_PATTERN
        
        # [FIX] Aggressively hide any active suggestion window
        filter_widgets = [
            getattr(self, 'cb_daily_filter_site', None),
            getattr(self, 'cb_daily_filter_material', None),
            getattr(self, 'cb_daily_filter_equipment', None),
            getattr(self, 'cb_daily_filter_worker', None),
            getattr(self, 'cb_daily_filter_vehicle', None)
        ]
        for widget in filter_widgets:
            if widget and hasattr(widget, '_suggestion_win'):
                widget._suggestion_win.hide()

        # [NEW] Move focus to tree to ensure keyboard/window state transition
        if hasattr(self, 'daily_usage_tree'):
            self.daily_usage_tree.focus_set()

        # Clear current view
        for item in self.daily_usage_tree.get_children():
            self.daily_usage_tree.delete(item)
        
        # Get filter values
        start_date_str = self.ent_daily_start_date.get().strip()
        end_date_str = self.ent_daily_end_date.get().strip()
        filter_site = self.cb_daily_filter_site.get().strip() if hasattr(self, 'cb_daily_filter_site') else '전체'
        filter_material = self.cb_daily_filter_material.get().strip() if hasattr(self, 'cb_daily_filter_material') else '전체'
        filter_equipment = self.cb_daily_filter_equipment.get().strip() if hasattr(self, 'cb_daily_filter_equipment') else '전체'
        filter_worker = self.cb_daily_filter_worker.get().strip() if hasattr(self, 'cb_daily_filter_worker') else '전체'
        filter_vehicle = self.cb_daily_filter_vehicle.get().strip() if hasattr(self, 'cb_daily_filter_vehicle') else '전체'
        filter_shift = self.cb_daily_filter_shift.get().strip() if hasattr(self, 'cb_daily_filter_shift') else '전체'
        
        # Ensure default values if empty
        if not filter_site: filter_site = '전체'
        if not filter_material: filter_material = '전체'
        if not filter_equipment: filter_equipment = '전체'
        if not filter_worker: filter_worker = '전체'
        if not filter_shift: filter_shift = '전체'
        
        # (Dropdown population moved to refresh_inquiry_filters for stability)
        
        # Parse dates
        try:
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None
            if end_date:
                # Include the entire end date
                end_date = end_date + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)
        except ValueError:
            messagebox.showwarning("입력 오류", "날짜 형식이 올바르지 않습니다. (YYYY-MM-DD)")
            return
        
        # Filter data
        filtered_df = self.daily_usage_df.copy()
        
        if start_date is not None:
            filtered_df = filtered_df[pd.to_datetime(filtered_df['Date']) >= start_date]
        
        if end_date is not None:
            filtered_df = filtered_df[pd.to_datetime(filtered_df['Date']) <= end_date]
        
        if filter_site != '전체':
            # Support partial match for site for more robustness since it's now editable
            filtered_df = filtered_df[filtered_df['Site'].astype(str).str.contains(filter_site, na=False, case=False, regex=False)]
        
        if filter_equipment != '전체':
            # Support partial match for equipment
            filtered_df = filtered_df[filtered_df['장비명'].astype(str).str.contains(filter_equipment, na=False, case=False, regex=False)]
        
        if filter_vehicle != '전체' and '차량번호' in filtered_df.columns:
            # Support partial match for vehicle number
            filtered_df = filtered_df[filtered_df['차량번호'].astype(str).str.contains(filter_vehicle, na=False, case=False, regex=False)]
        
        # Filter by material if specified
        if filter_material != '전체':
            # Check all materials for matching full display name
            matching_mat_ids = []
            if not self.materials_df.empty:
                for _, mat in self.materials_df.iterrows():
                    if self.get_material_display_name(mat['MaterialID']) == filter_material:
                        matching_mat_ids.append(mat['MaterialID'])
            
            if matching_mat_ids:
                filtered_df = filtered_df[filtered_df['MaterialID'].isin(matching_mat_ids)]
            else:
                # If no direct match in materials_df, try basic 품목명 (for edge cases)
                matching_mat_ids = self.materials_df[self.materials_df['품목명'] == filter_material]['MaterialID'].tolist()
                if matching_mat_ids:
                    filtered_df = filtered_df[filtered_df['MaterialID'].isin(matching_mat_ids)]

        # Filter by worker or shift (Smarter & Marker-Insensitive)
        if filter_worker != '전체' or filter_shift != '전체':
            worker_cols = ['User', 'User2', 'User3', 'User4', 'User5', 'User6', 'User7', 'User8', 'User9', 'User10']
            time_cols = ['WorkTime', 'WorkTime2', 'WorkTime3', 'WorkTime4', 'WorkTime5', 'WorkTime6', 'WorkTime7', 'WorkTime8', 'WorkTime9', 'WorkTime10']
            
            # Normalize filter text: remove spaces and lowercase for maximum flexibility
            fw_clean = filter_worker.replace(' ', '').lower()
            if fw_clean == '': fw_clean = '전체' # Handle empty input as showing all

            def row_matches(row):
                if fw_clean == '전체' and filter_shift == '전체':
                    return True
                
                # Pre-clean filter text for comparison
                f_worker = fw_clean
                f_shift = f"({filter_shift})" if filter_shift != '전체' else None
                
                for i in range(len(worker_cols)):
                    # Securely get worker and time data
                    w_col = worker_cols[i]
                    t_col = time_cols[i]
                    
                    if w_col not in row: continue
                    w_val_raw = str(row[w_col]).strip()
                    
                    if not w_val_raw or w_val_raw.lower() in ['nan', '0.0', 'none', '']:
                        continue
                        
                    t_val = str(row.get(t_col, '')).strip()
                    
                    # Worker match (Substring match on cleaned names)
                    w_match = True
                    if f_worker != '전체':
                        w_val_clean = w_val_raw.replace(' ', '').lower()
                        # Also handle records where marker is still in User column
                        w_val_clean = marker_pattern.sub('', w_val_clean).strip()
                        w_match = f_worker in w_val_clean
                    
                    # Shift match (Look for marker in WorkTime column)
                    s_match = True
                    if f_shift:
                        s_match = f_shift in t_val
                    
                    if w_match and s_match:
                        return True
                return False

            # Apply robust filtering
            filtered_df = filtered_df[filtered_df.apply(row_matches, axis=1)]
        
        # Sort by date descending - ensure Date column is consistent type
        if not filtered_df.empty:
            # Convert Date column to datetime for consistent sorting
            filtered_df['Date'] = pd.to_datetime(filtered_df['Date'], errors='coerce')
            filtered_df = filtered_df.sort_values('Date', ascending=False)
        
        # Define RTK categories
        rtk_categories = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타", "총계"]
        
        # Display entries and calculate totals
        # Display entries and calculate totals
        total_film_count = 0
        total_rtk = [0.0] * len(rtk_categories)
        total_ndt = [0.0] * 7
        total_test_amount = 0.0
        total_unit_price = 0.0
        total_travel_cost = 0.0
        total_meal_cost = 0.0
        total_test_fee = 0.0
        total_ot_hours = 0.0
        total_ot_amount = 0
        total_work_hours = 0.0 # Added for Total Working Time
        total_mileage = 0.0    # Sum of individual entries if applicable
        min_mileage = float('inf')
        max_mileage = float('-inf')
        total_indiv_ot_hours = [0.0] * 10
        total_indiv_ot_amounts = [0] * 10
        
        has_equip = False
        has_method = False
        has_note = False
        
        
        current_date = None
        

        for idx, entry in filtered_df.iterrows():
            usage_date = self._safe_format_datetime(entry.get('Date', ''), '%Y-%m-%d')
            if not usage_date:
                usage_date = "Unknown"
            
            current_date = usage_date
            
            mat_id = entry['MaterialID']
            mat_name = self.get_material_display_name(mat_id)
            
            entry_time = self._safe_format_datetime(entry.get('EntryTime', ''), '%Y-%m-%d %H:%M:%S')
            
            # Consolidate workers
            def clean_str(val):
                return self.clean_nan(val)

            all_users = [
                clean_str(entry.get('User', '')),
                clean_str(entry.get('User2', '')),
                clean_str(entry.get('User3', '')),
                clean_str(entry.get('User4', '')),
                clean_str(entry.get('User5', '')),
                clean_str(entry.get('User6', '')),
                clean_str(entry.get('User7', '')),
                clean_str(entry.get('User8', '')),
                clean_str(entry.get('User9', '')),
                clean_str(entry.get('User10', ''))
            ]
            
            # [FILTER FIX] If filtering by worker, show only that worker's name and THEIR worktime
            display_worktime = clean_str(entry.get('WorkTime', '')) # Default to User1
            raw_workers = [u for u in all_users if u]
            
            if filter_worker != '전체':
                consolidated_workers = filter_worker
                
                # Find the worktime for this specific worker
                f_worker_clean = filter_worker.replace(' ', '').lower()
                for i in range(1, 11):
                    u_key = 'User' if i == 1 else f'User{i}'
                    if u_key not in entry: continue
                    u_val_raw = clean_str(entry.get(u_key, ''))
                    u_val_clean = marker_pattern.sub('', u_val_raw.replace(' ', '').lower()).strip()
                    
                    if u_val_clean == f_worker_clean:
                        wt_key = 'WorkTime' if i == 1 else f'WorkTime{i}'
                        display_worktime = clean_str(entry.get(wt_key, ''))
                        break
            else:
                consolidated_workers = self.format_worker_summary(raw_workers)
            
            # Get film count - consistently use FilmCount
            film_count = entry.get('FilmCount', 0)
            if pd.notna(film_count):
                film_count_val = float(film_count)
                total_film_count += film_count_val
                film_count_str = f"{film_count_val:.1f}"
            else:
                film_count_str = "0.0"
            
            # Accumulate cost totals
            # [NEW] Ensure numeric types for calculation and formatting
            def to_f(val):
                try:
                    if pd.isna(val) or str(val).lower() == 'nan': return 0.0
                    return float(str(val).replace(',', ''))
                except: return 0.0

            q_val = to_f(entry.get('검사량', 0.0))
            p_val = to_f(entry.get('단가', 0.0))
            t_val_cost = to_f(entry.get('출장비', 0.0))
            m_val_cost = to_f(entry.get('일식', 0.0))
            f_val_cost = to_f(entry.get('검사비', 0.0))

            total_test_amount += q_val
            total_unit_price += p_val
            total_travel_cost += t_val_cost
            total_meal_cost += m_val_cost
            total_test_fee += f_val_cost

            # Cumulative mileage
            m_val = to_f(entry.get('주행거리', 0))
            total_mileage += m_val
            if m_val > 0:
                min_mileage = min(min_mileage, m_val)
                max_mileage = max(max_mileage, m_val)
            
            if clean_str(entry.get('장비명', '')): has_equip = True
            if clean_str(entry.get('검사방법', '')): has_method = True
            if clean_str(entry.get('Note', '')): has_note = True
            
            # Calculate row OT total
            row_ot_hours = 0.0
            row_ot_amount = 0
            row_ots = []
            
            # Calculate row-specific work duration for total accumulation
            row_duration = 0.0
            
            # Use the already determined display_worktime for the "Total" of this column
            # This ensures the total below the column matches the sum of values visible in that column.
            if display_worktime and '~' in display_worktime:
                try:
                    # Clean off markers like (주간) before splitting
                    clean_wt = marker_pattern.sub('', display_worktime).strip()
                    if '~' in clean_wt:
                        start_str, end_str = clean_wt.split('~')
                        sh, sm = map(int, start_str.split(':'))
                        eh, em = map(int, end_str.split(':'))
                        
                        start_min = sh * 60 + sm
                        end_min = eh * 60 + em
                        if end_min < start_min: end_min += 24 * 60
                        
                        row_duration = (end_min - start_min) / 60.0
                        total_work_hours += row_duration
                except:
                    pass

            for i in range(1, 11):
                # Check worker filter for this slot
                user_key = 'User' if i == 1 else f'User{i}'
                user_val = self.clean_nan(entry.get(user_key, ''))
                
                if filter_worker != '전체' and user_val != filter_worker:
                    row_ots.append("") # Empty string looks cleaner than "0" for hidden cols
                    continue

                ot_str = clean_str(entry.get(f'OT{i}' if i > 1 else 'OT', ''))
                if ot_str:
                    try:
                        # [NEW] Robust parsing for both formats
                        if '(' in ot_str and '원)' in ot_str:
                            h_part = float(ot_str.split('시간')[0])
                            a_part = int(ot_str.split('(')[1].split('원')[0].replace(',', ''))
                        elif ot_str.replace(',', '').isdigit():
                            a_part = int(ot_str.replace(',', ''))
                            # Calculate hours from worktime if amount only
                            wt_key = 'WorkTime' if i == 1 else f'WorkTime{i}'
                            wt_val = clean_str(entry.get(wt_key, ''))
                            h_part, _ = self._calculate_ot_from_worktime(wt_val, pd.to_datetime(entry['Date']))
                        else:
                            a_part = 0
                            h_part = self._parse_ot_hours(ot_str)
                            
                        row_ot_hours += h_part
                        row_ot_amount += a_part
                        total_indiv_ot_hours[i-1] += h_part
                        total_indiv_ot_amounts[i-1] += a_part
                        row_ots.append(f"{a_part:,}")
                    except:
                        row_ots.append(ot_str)
                else:
                    row_ots.append("0")
            
            total_ot_hours += row_ot_hours
            total_ot_amount += row_ot_amount
            ot_sum_display = f"{row_ot_hours:.1f}시간" if row_ot_hours > 0 else "0"
            if row_ot_amount > 0:
                ot_sum_display += f" ({row_ot_amount:,}원)"
            
            # Individual OT values for columns (Simplified: Amount only)
            # row_ots is already built above!
            
            # Get RTK values
            rtk_values = []
            row_rtk_total = 0
            # Use categories without '총계' for individual columns
            rtk_cats_only = rtk_categories[:-1]
            
            for i, category in enumerate(rtk_cats_only):
                value = entry.get(f'RTK_{category}', 0)
                if pd.notna(value):
                    val_float = float(value)
                    rtk_values.append(f"{val_float:.1f}")
                    row_rtk_total += val_float
                    total_rtk[i] += val_float
                else:
                    rtk_values.append("0.0")
            
            rtk_values.append(f"{row_rtk_total:.1f}")  # Add row total
            total_rtk[7] += row_rtk_total
            
            # Get NDT materials values
            ndt_values = []
            ndt_materials = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
            
            for i, material in enumerate(ndt_materials):
                # Map new names to old keys for backward compatibility if needed
                col_name = f'NDT_{material}'
                value = entry.get(col_name)
                # Fallback for old data
                if pd.isna(value):
                    if material == "흑색자분": value = entry.get('NDT_자분', 0)
                    elif material == "백색페인트": value = entry.get('NDT_페인트', 0)
                    elif material == "형광침투제": value = entry.get('NDT_형광', 0)
                    else: value = 0
                if pd.notna(value):
                    val_float = float(value)
                    ndt_values.append(f"{val_float:.1f}")
                    total_ndt[i] += val_float
                else:
                    ndt_values.append("0.0")
            
            # Get remark with manager name if exists
            user_val = entry.get('User', '')
            if pd.isna(user_val) or str(user_val).lower() == 'nan': user_val = ''
            
            note_val = entry.get('Note', '')
            if pd.isna(note_val) or str(note_val).lower() == 'nan': note_val = ''
            
            display_note = f"[{user_val}] {note_val}" if user_val else note_val
            
            # Insert with index as tag for reliable deletion
            # def clean_str(val): -> Removed, confirmed defined above
            #    return str(val).replace('nan', '').replace('None', '').strip()

            self.daily_usage_tree.insert('', tk.END, values=(
                usage_date,
                entry.get('Site', ''),
                consolidated_workers,
                display_worktime,
                *row_ots,
                entry.get('장비명', ''),
                entry.get('검사방법', ''),
                f"{q_val:.1f}",
                film_count_str,
                f"{p_val:,.0f}",
                f"{t_val_cost:,.0f}",
                f"{m_val_cost:,.0f}",
                f"{f_val_cost:,.0f}",
                f"{row_ot_hours:.1f}", # OT시간
                f"{row_ot_amount:,}",  # OT금액
                mat_name,
                *rtk_values,
                *ndt_values,
                display_note if not note_val else f"{display_note}", 
                entry_time,
                clean_str(entry.get('차량번호', '')),
                clean_str(entry.get('주행거리', '')),
                clean_str(entry.get('차량점검', '')),
                clean_str(entry.get('차량비고', '')),
                ", ".join(raw_workers) # (Full작업자) storage for Excel export
            ), tags=(str(idx),))
            
        # Insert last daily subtotal and final total row if data exists
        if not filtered_df.empty:
            
            # Final overall total
            self.daily_usage_tree.tag_configure('total', background='#E8F4F8', font=('Arial', 9, 'bold'))
            
            total_values = [
                '--- 전체 누계 ---',
                '', # 현장
                '', # 작업자
                f"{total_work_hours:.1f} Hrs", # 작업시간 (Calculated Total)
                # Individual OT Totals (Simplified: Amount only)
                *[f"{a:,}" if a > 0 else "0" for a in total_indiv_ot_amounts],
                '', # 장비명
                '', # 검사방법
                f"{total_test_amount:.1f}",
                f"{total_film_count:.1f}",
                f"{total_unit_price:,.0f}",
                f"{total_travel_cost:,.0f}",
                f"{total_meal_cost:,.0f}",
                f"{total_test_fee:,.0f}",
                f"{total_ot_hours:.1f}", # OT시간 합계
                f"{total_ot_amount:,}",  # OT금액 합계
                '', # 품목명
                *[f"{v:.1f}" for v in total_rtk],
                *[f"{v:.1f}" for v in total_ndt],
                '',   # 비고
                '',   # 입력시간
                '',   # 차량번호
                f"이동: {max_mileage - min_mileage:,.1f} km" if max_mileage > min_mileage else "0 km",   # 주행거리 차이
                '',   # 차량점검
                '',   # 차량비고
                ''    # (Full작업자)
            ]
            self.daily_usage_tree.tag_configure('total', background='#E8F4F8', font=('Arial', 12, 'bold'))
            self.daily_usage_tree.insert('', tk.END, values=total_values, tags=('total',))
            
            # --- Dynamic Column Hiding ---
            # Mandatory cols are those that are ALWAYS essential (Site, Worker, Material, EntryTime, Date)
            mandatory_cols = ['날짜', '현장', '작업자', '품목명', '입력시간']
            
            # Map dynamic columns to their total values
            dynamic_col_status = {
                '작업시간': total_work_hours > 0,
                '수량': total_test_amount > 0,
                '필름매수': total_film_count > 0,
                '단가': total_unit_price > 0,
                '출장비': total_travel_cost > 0,
                '일식': total_meal_cost > 0,
                '검사비': total_test_fee > 0,
                'OT시간': total_ot_hours > 0,
                'OT금액': total_ot_amount > 0,
                '장비명': has_equip,
                '검사방법': has_method,
                '비고': has_note,
                '차량번호': not filtered_df.empty and (filtered_df['차량번호'].astype(str).str.strip() != '').any(),
                '주행거리': not filtered_df.empty and (filtered_df['주행거리'].astype(str).str.strip() != '').any(),
                '차량점검': not filtered_df.empty and (filtered_df['차량점검'].astype(str).str.strip() != '').any(),
                '차량비고': not filtered_df.empty and (filtered_df['차량비고'].astype(str).str.strip() != '').any()
            }
            
            # Individual OT columns
            for i in range(1, 11):
                col_name = f'OT{i}'
                dynamic_col_status[col_name] = total_indiv_ot_amounts[i-1] > 0
            
            # RTK Columns (11-17) and RTK총계 (18)
            rtk_col_names = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타"]
            for i, col_name in enumerate(rtk_col_names):
                dynamic_col_status[col_name] = total_rtk[i] > 0
            dynamic_col_status['RTK총계'] = total_rtk[7] > 0
            
            # NDT Columns (19-25)
            ndt_col_names = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
            for i, col_name in enumerate(ndt_col_names):
                dynamic_col_status[col_name] = total_ndt[i] > 0
                
            # Build final visible columns list based on the order in self.daily_usage_tree['columns']
            all_cols = self.daily_usage_tree['columns']
            
            # If user has manually selected columns, respect those selections
            if hasattr(self, 'manual_visible_cols') and self.manual_visible_cols:
                # Use manual selection directly but filter for columns that actually exist
                visible_cols = [col for col in self.manual_visible_cols if col in all_cols]
                
                # [OPTIONAL] If some column NOT in manual_visible_cols has data (>0), we can choose to auto-show it 
                # to prevent data loss perception, or strictly follow manual. 
                # Let's prioritize manual, but still ALWAYS include mandatory cols.
                for m_col in mandatory_cols:
                    if m_col in all_cols and m_col not in visible_cols:
                        visible_cols.append(m_col) # Simplified insert, order fixed below
                
                # Re-sort visible_cols based on original tree columns order
                visible_cols = [col for col in all_cols if col in visible_cols]
            else:
                # Default: Show all defined columns except internal technical ones
                visible_cols = [col for col in all_cols if col != '(Full작업자)']
            
            # Apply visible columns to treeview
            self.daily_usage_tree['displaycolumns'] = visible_cols
            
            # [REMOVED] Header rename '필름매수' to 'RTK' per user request

            # Ensure stretch=False and minwidth is relaxed for ALL displayed columns
            # This is critical for making all columns, especially the last one, resizable
            for col in visible_cols:
                self.daily_usage_tree.column(col, stretch=False, minwidth=20)

            # Re-apply saved column widths if available
            if hasattr(self, 'tab_config') and 'daily_usage_col_widths' in self.tab_config:
                saved_widths = self.tab_config['daily_usage_col_widths']
                for col in visible_cols:
                    if col in saved_widths:
                        try:
                            self.daily_usage_tree.column(col, width=int(saved_widths[col]), stretch=False)
                            # Enforce minimums for high-precision cols to prevent truncation
                            if col == '날짜':
                                self.daily_usage_tree.column(col, width=max(int(saved_widths[col]), 160), stretch=False)
                            elif col == '입력시간':
                                self.daily_usage_tree.column(col, width=max(int(saved_widths[col]), 300), stretch=False)
                        except: pass
            
            # Ensure Total Row stays at bottom
            self.daily_usage_tree.detach(self.daily_usage_tree.get_children()[-1])
            self.daily_usage_tree.insert('', tk.END, values=total_values, tags=('total',))
        else:
            # If empty, check manual visibility first
            if hasattr(self, 'manual_visible_cols') and self.manual_visible_cols:
                visible_cols = [col for col in self.manual_visible_cols if col in self.daily_usage_tree['columns']]
                self.daily_usage_tree['displaycolumns'] = visible_cols
            else:
                # Default: Show all columns
                visible_cols = [col for col in self.daily_usage_tree['columns'] if col != '(Full작업자)']
                self.daily_usage_tree['displaycolumns'] = visible_cols
            
            # Ensure stretch=False and minwidth is relaxed for all displayed columns
            for col in visible_cols:
                self.daily_usage_tree.column(col, stretch=False, minwidth=20)

            # Re-apply saved column widths if available
            if hasattr(self, 'tab_config') and 'daily_usage_col_widths' in self.tab_config:
                saved_widths = self.tab_config['daily_usage_col_widths']
                for col in visible_cols:
                    if col in saved_widths:
                        try:
                            self.daily_usage_tree.column(col, width=int(saved_widths[col]), stretch=False)
                            # Enforce minimums for high-precision cols
                            if col == '날짜':
                                self.daily_usage_tree.column(col, width=max(int(saved_widths[col]), 160), stretch=False)
                            elif col == '입력시간':
                                self.daily_usage_tree.column(col, width=max(int(saved_widths[col]), 300), stretch=False)
                        except: pass
    
    def _auto_adjust_tree_columns(self, tree, expand_only=False):
        """Automatically adjust column widths to fit content"""
        import tkinter.font as tkfont
        
        # Use the actual font used in the Treeview (12pt Malgun Gothic)
        # This ensures measurement matches display perfectly.
        content_font = tkfont.Font(family="Malgun Gothic", size=12)
        heading_font = tkfont.Font(family="Malgun Gothic", size=12, weight="bold")
        
        # We only care about visible columns
        visible_cols = tree['displaycolumns']
        if not visible_cols or visible_cols == ('#all'):
             visible_cols = tree['columns']

        for col in visible_cols:
            # Maximum Overkill: Character Count Heuristic
            # Assume 40px per character + 200px buffer
            # This GUARANTEES visibility in any environment
            
            # Measure heading
            w = len(col) * 40 + 200
            
            # Measure content
            col_index = list(tree['columns']).index(col)
            for item in tree.get_children():
                val = tree.item(item, 'values')
                if val and col_index < len(val):
                    text_len = len(str(val[col_index]))
                    content_w = text_len * 40 + 200
                    if content_w > w:
                        w = content_w
            
            # Apply specific constraints
            # Ultra-Safe: No max limit for potentially long text columns
            if col in ['품목명', '작업자', '비고']:
                w = max(w, 300) # Minimum base
                if col == '품목명': w = max(w, 700)
                elif col == '작업자': w = max(w, 300)
                elif col == '비고': w = max(w, 600)
            else:
                # Other columns get a relaxed global max
                min_w = 200
                if col == '날짜': min_w = 400
                elif col == '현장': min_w = 300
                elif col == '장비명': min_w = 250
                elif col == '검사방법': min_w = 200
                
                w = max(min_w, min(w, 2000)) 
            
            # Smart Auto-Expand: Only grow if content requires it
            if expand_only:
                current_w = int(tree.column(col, 'width'))
                if w <= current_w:
                    continue  # Keep user's manual adjustment
            
            # All columns are now user-resizable
            tree.column(col, width=w, minwidth=20, stretch=False, anchor='center')
        
        # Force layout update
        tree.update_idletasks()

    def delete_daily_usage_entry(self):
        """선택된 일일 사용량 기록 삭제 (인덱스 기반으로 정확하게 삭제 및 재고 환원)"""
        selected_items = self.daily_usage_tree.selection()
        
        if not selected_items:
            messagebox.showwarning("선택 오류", "삭제할 항목을 선택해주세요.")
            return
        
        # 사용자 확인
        result = messagebox.askyesno("삭제 확인", f"{len(selected_items)}개의 기록을 삭제하시겠습니까?\n(삭제 시 차감되었던 재고도 자동으로 환원됩니다.)")
        
        if not result:
            return
        
        indices_to_delete = []
        for item in selected_items:
            tags = self.daily_usage_tree.item(item, 'tags')
            if tags:
                try:
                    df_idx = int(tags[0])
                    indices_to_delete.append(df_idx)
                except ValueError:
                    continue
        
        if not indices_to_delete:
            messagebox.showwarning("삭제 실패", "삭제할 항목의 데이터 정보를 찾을 수 없습니다.")
            return

        deleted_count = 0
        try:
            # Sort indices in descending order to avoid index shift issues during deletion
            # But wait, we are using boolean masking or actual indices in the original dataframe
            # Since we will drop them at once or one by one from the original, it's better to collect all indices.
            
            for idx in indices_to_delete:
                if idx not in self.daily_usage_df.index:
                    continue
                
                entry = self.daily_usage_df.loc[idx]
                
                # --- 재고 환원 로직 (해당 내역으로 인해 차감되었던 트랜잭션 삭제) ---
                site = entry.get('Site', '')
                usage_date = pd.to_datetime(entry.get('Date'))
                mat_id = entry.get('MaterialID')
                
                # 해당 내역과 관련된 트랜잭션 찾기 (날짜, MaterialID, Note가 현장 사용 자동 차감인 것)
                # Note 형식을 add_daily_usage_entry와 동일하게 구성
                note_pattern = f'{site} 현장 사용 (자동 차감)'
                
                # 필터링하여 관련 트랜잭션 삭제
                if not self.transactions_df.empty:
                    trans_mask = (
                        (pd.to_datetime(self.transactions_df['Date']).dt.date == usage_date.date()) &
                        (self.transactions_df['MaterialID'] == mat_id) &
                        (self.transactions_df['Type'] == 'OUT') &
                        (self.transactions_df['Note'] == note_pattern)
                    )
                    
                    # NDT 자재 트랜잭션도 포함될 수 있음. NDT 자재들은 MaterialID가 다를 수 있으나 
                    # Note는 동일하게 "현장 사용 (자동 차감)"으로 들어감.
                    # NDT 자재들의 MaterialID는 entry에 NDT_xxx 필드로 명시되어 있음
                    ndt_materials = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
                    ndt_mat_ids = []
                    for ndt_name in ndt_materials:
                        if entry.get(f'NDT_{ndt_name}', 0) > 0:
                            ndt_mat_rows = self.materials_df[self.materials_df['품목명'] == ndt_name]
                            if not ndt_mat_rows.empty:
                                ndt_mat_ids.append(ndt_mat_rows['MaterialID'].values[0])
                    
                    if ndt_mat_ids:
                        trans_mask = trans_mask | (
                            (pd.to_datetime(self.transactions_df['Date']).dt.date == usage_date.date()) &
                            (self.transactions_df['MaterialID'].isin(ndt_mat_ids)) &
                            (self.transactions_df['Type'] == 'OUT') &
                            (self.transactions_df['Note'] == note_pattern)
                        )

                    # 트랜잭션 삭제
                    self.transactions_df = self.transactions_df[~trans_mask]
                    deleted_count += 1

            # 기록 삭제
            self.daily_usage_df = self.daily_usage_df.drop(indices_to_delete)
            
            # 인덱스 초기화는 하지 않음 (태그 매칭을 위해 원본 인덱스 유지 권장하나, 
            # drop 후에는 뷰 업데이트 시 어차피 다시 생성되므로 안전하게 reset_index 가능)
            self.daily_usage_df = self.daily_usage_df.reset_index(drop=True)
            
            if self.save_data():
                messagebox.showinfo("삭제 완료", f"{len(indices_to_delete)}개의 기록이 삭제되고 재고가 환원되었습니다.")
                self.update_daily_usage_view()
                self.update_stock_view()
                self.update_transaction_view()
                self.refresh_inquiry_filters()
        except Exception as e:
            messagebox.showerror("삭제 오류", f"기록 삭제 중 오류가 발생했습니다: {e}")

    def open_edit_daily_usage_dialog(self):
        """Open a dialog to edit the selected daily usage record"""
        selection = self.daily_usage_tree.selection()
        if not selection:
            messagebox.showwarning("선택 오류", "수정할 항목을 선택해주세요.")
            return
            
        item = self.daily_usage_tree.item(selection[0])
        tags = item.get('tags', [])
        if not tags:
            messagebox.showwarning("데이터 오류", "선택한 항목의 데이터 정보를 찾을 수 없습니다.")
            return
            
        try:
            df_idx = int(tags[0])
            entry_data = self.daily_usage_df.loc[df_idx]
        except (ValueError, KeyError, IndexError):
            messagebox.showwarning("데이터 오류", "데이터를 불러오는 중 오류가 발생했습니다.")
            return

        # Create Edit Dialog
        edit_win = tk.Toplevel(self.root)
        edit_win.title("일일 사용 기록 수정")
        edit_win.geometry("800x900")
        edit_win.transient(self.root)
        edit_win.grab_set()

        # Main scrollable frame
        container = ttk.Frame(edit_win)
        container.pack(expand=True, fill='both')
        
        canvas = tk.Canvas(container)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas, padding=20)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Scroll binding is handled globally in MaterialManager.__init__

        fields = {}
        
        # 1. Basic Info
        basic_frame = ttk.LabelFrame(scrollable_frame, text="기본 정보", padding=10)
        basic_frame.pack(fill='x', pady=5)
        
        # Date
        ttk.Label(basic_frame, text="날짜:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        ent_date = DateEntry(basic_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd', locale='ko_KR', state='readonly', showweeknumbers=True)
        curr_date = entry_data.get('Date', datetime.datetime.now().strftime('%Y-%m-%d'))
        try: ent_date.set_date(datetime.datetime.strptime(str(curr_date), '%Y-%m-%d'))
        except: pass
        ent_date.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        fields['Date'] = ent_date

        # Site
        ttk.Label(basic_frame, text="현장:").grid(row=0, column=2, padx=5, pady=5, sticky='w')
        cb_site = ttk.Combobox(basic_frame, width=20, values=self.sites)
        cb_site.set(self.clean_nan(entry_data.get('Site', '')))
        cb_site.grid(row=0, column=3, padx=5, pady=5, sticky='w')
        fields['Site'] = cb_site
        
        # Material
        ttk.Label(basic_frame, text="품목명:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
        cb_mat = ttk.Combobox(basic_frame, width=40)
        # Material list reconstruction
        mat_options = []
        for _, m_row in self.materials_df.iterrows():
            m_name = str(m_row.get('품목명', ''))
            m_model = self.clean_nan(m_row.get('모델명', ''))
            m_sn = self.clean_nan(m_row.get('SN', ''))
            m_display = m_name
            if m_model: m_display += f" - {m_model}"
            if m_sn: m_display += f" (SN: {m_sn})"
            mat_options.append(m_display)
        cb_mat['values'] = mat_options
        
        # Determine current display name
        curr_mat_id = entry_data.get('MaterialID')
        curr_mat_display = ""
        m_rows = self.materials_df[self.materials_df['MaterialID'] == curr_mat_id]
        if not m_rows.empty:
            m_row = m_rows.iloc[0]
            m_name = str(m_row.get('품목명', ''))
            m_model = self.clean_nan(m_row.get('모델명', ''))
            m_sn = self.clean_nan(m_row.get('SN', ''))
            curr_mat_display = m_name
            if m_model: curr_mat_display += f" - {m_model}"
            if m_sn: curr_mat_display += f" (SN: {m_sn})"
        
        cb_mat.set(curr_mat_display)
        cb_mat.grid(row=1, column=1, columnspan=3, padx=5, pady=5, sticky='w')
        fields['Material'] = cb_mat

        # Equipment & Method
        ttk.Label(basic_frame, text="장비명:").grid(row=2, column=0, padx=5, pady=5, sticky='w')
        cb_equip = ttk.Combobox(basic_frame, width=20, values=self.equipments)
        cb_equip.set(self.clean_nan(entry_data.get('장비명', '')))
        cb_equip.grid(row=2, column=1, padx=5, pady=5, sticky='w')
        fields['장비명'] = cb_equip

        ttk.Label(basic_frame, text="검사방법:").grid(row=2, column=2, padx=5, pady=5, sticky='w')
        cb_method = ttk.Combobox(basic_frame, width=10, values=['RT', 'PAUT', 'UT', 'MT', 'PT', 'ETC'])
        cb_method.set(self.clean_nan(entry_data.get('검사방법', '')))
        cb_method.grid(row=2, column=3, padx=5, pady=5, sticky='w')
        fields['검사방법'] = cb_method

        # 2. Quantities & Costs
        cost_frame = ttk.LabelFrame(scrollable_frame, text="수량 및 비용", padding=10)
        cost_frame.pack(fill='x', pady=5)
        
        qty_configs = [
            ('검사량', '검사량'), ('필름매수', 'FilmCount'), ('단가', '단가'), ('출장비', '출장비'),
            ('일식', '일식'), ('검사비', '검사비')
        ]
        for i, (lbl, key) in enumerate(qty_configs):
            row = i // 3
            col = (i % 3) * 2
            ttk.Label(cost_frame, text=f"{lbl}:").grid(row=row, column=col, padx=5, pady=5, sticky='w')
            ent = ttk.Entry(cost_frame, width=12)
            val = entry_data.get(key, 0)
            if pd.isna(val): val = 0
            ent.insert(0, str(val))
            ent.grid(row=row, column=col+1, padx=5, pady=5, sticky='w')
            fields[key] = ent

        # 3. RTK & NDT
        material_frame = ttk.Frame(scrollable_frame)
        material_frame.pack(fill='x', pady=5)
        
        rtk_frame = ttk.LabelFrame(material_frame, text="RTK 사용량 (재촬영)", padding=10)
        rtk_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))
        
        rtk_cats = ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타"]
        for i, cat in enumerate(rtk_cats):
            ttk.Label(rtk_frame, text=f"{cat}:").grid(row=i, column=0, padx=5, pady=2, sticky='w')
            ent = ttk.Entry(rtk_frame, width=10)
            val = entry_data.get(f'RTK_{cat}', 0)
            if pd.isna(val): val = 0
            ent.insert(0, str(val))
            ent.grid(row=i, column=1, padx=5, pady=2, sticky='w')
            fields[f'RTK_{cat}'] = ent

        ndt_frame = ttk.LabelFrame(material_frame, text="NDT 약품 사용량", padding=10)
        ndt_frame.pack(side='left', fill='both', expand=True, padx=(5, 0))
        
        ndt_cats = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
        for i, cat in enumerate(ndt_cats):
            ttk.Label(ndt_frame, text=f"{cat}:").grid(row=i, column=0, padx=5, pady=2, sticky='w')
            ent = ttk.Entry(ndt_frame, width=10)
            m_key = "".join(cat.split())
            val = entry_data.get(f'NDT_{m_key}', 0)
            if pd.isna(val): val = 0
            ent.insert(0, str(val))
            ent.grid(row=i, column=1, padx=5, pady=2, sticky='w')
            fields[f'NDT_{m_key}'] = ent

            fields[f'NDT_{m_key}'] = ent
            
        # 4. Vehicle Info
        vehicle_frame = ttk.LabelFrame(scrollable_frame, text="차량 점검 정보", padding=10)
        vehicle_frame.pack(fill='x', pady=5)
        
        v_configs = [
            ('차량번호', '차량번호'), ('주행거리', '주행거리'),
            ('점검내용', '차량점검'), ('차량비고', '차량비고')
        ]
        for v_i, (v_lbl, v_key) in enumerate(v_configs):
            v_row = v_i // 2
            v_col = (v_i % 2) * 2
            ttk.Label(vehicle_frame, text=f"{v_lbl}:").grid(row=v_row, column=v_col, padx=5, pady=5, sticky='w')
            v_ent = ttk.Entry(vehicle_frame, width=25)
            v_val = entry_data.get(v_key, '')
            v_ent.insert(0, self.clean_nan(v_val))
            v_ent.grid(row=v_row, column=v_col+1, padx=5, pady=5, sticky='w')
            fields[v_key] = v_ent

        # 5. Workers
        worker_main_frame = ttk.LabelFrame(scrollable_frame, text="작업자 정보 (1~10)", padding=10)
        worker_main_frame.pack(fill='x', pady=5)
        
        # [NEW] Inner-dialog helper for RT film sync
        def sync_edit_film(event=None):
            method = cb_method.get().strip().upper()
            if "RT" == method or method.startswith("RT "):
                qty = fields['검사량'].get().strip()
                fields['FilmCount'].delete(0, tk.END)
                fields['FilmCount'].insert(0, qty)

        # [NEW] Inner-dialog helper for Fee calculation
        def update_edit_fee_calc(event=None):
            try:
                def get_f(val):
                    try:
                        v = str(val).strip().replace(',', '')
                        return float(v) if v else 0.0
                    except: return 0.0

                qty = get_f(fields['검사량'].get())
                price = get_f(fields['단가'].get())
                travel = get_f(fields['출장비'].get())
                meal = get_f(fields['일식'].get())
                
                calc_fee = (qty * price) + travel + meal
                fields['검사비'].delete(0, tk.END)
                fields['검사비'].insert(0, f"{calc_fee:,.0f}")
            except: pass

        # Bindings for auto-sync and calc
        cb_method.bind('<<ComboboxSelected>>', sync_edit_film)
        fields['검사량'].bind('<KeyRelease>', lambda e: [sync_edit_film(), update_edit_fee_calc()])
        
        # [NEW] Add comma auto-formatting for edit fields
        for k in ['단가', '출장비', '일식', '검사비']:
            if k in fields:
                fields[k].bind('<KeyRelease>', update_edit_fee_calc, add='+')
                fields[k].bind('<FocusOut>', lambda e, widget=fields[k]: self.format_entry_with_commas(e, widget), add='+')
                fields[k].bind('<Return>', lambda e, widget=fields[k]: self.format_entry_with_commas(e, widget), add='+')

        # Grid for workers: 5 rows x 2 cols
        worker_fields = {}
        for i in range(1, 11):
            w_row = (i-1) // 2
            w_col = (i-1) % 2
            
            w_frame = ttk.Frame(worker_main_frame, padding=2)
            w_frame.grid(row=w_row, column=w_col, padx=10, pady=5, sticky='nw')
            
            ttk.Label(w_frame, text=f"{i}:", width=2).pack(side='left')
            
            # Name
            cb_name = ttk.Combobox(w_frame, width=8, values=self.users)
            name_key = 'User' if i == 1 else f'User{i}'
            cb_name.set(self.clean_nan(entry_data.get(name_key, '')))
            cb_name.pack(side='left', padx=2)
            worker_fields[f'name{i}'] = cb_name
            
            # Time
            cb_time = ttk.Combobox(w_frame, width=16, values=self.worktimes)
            time_key = 'WorkTime' if i == 1 else f'WorkTime{i}'
            cb_time.set(self.clean_nan(entry_data.get(time_key, '')))
            cb_time.pack(side='left', padx=2)
            worker_fields[f'time{i}'] = cb_time
            
            # OT
            ent_ot = ttk.Entry(w_frame, width=18)
            ot_key = 'OT' if i == 1 else f'OT{i}'
            ent_ot.insert(0, self.clean_nan(entry_data.get(ot_key, '')))
            ent_ot.pack(side='left', padx=2)
            worker_fields[f'ot{i}'] = ent_ot
            
            # Auto-calculate OT for each worker when time changes
            def on_time_changed(event, idx=i, w_ent=cb_time, ot_ent=ent_ot, d_ent=ent_date):
                self.calculate_and_update_ot_manual(w_ent.get(), ot_ent, d_ent.get_date())
            
            cb_time.bind('<Return>', on_time_changed)
        
        # [NEW] Apply autocomplete to main edit fields
        self.enable_autocomplete(cb_site, values_list_attr='sites')
        self.enable_autocomplete(cb_mat, values_list=mat_options)
        self.enable_autocomplete(cb_equip, values_list_attr='equipment_suggestions')

        # 5. Buttons
        btn_frame = ttk.Frame(scrollable_frame, padding=20)
        btn_frame.pack(fill='x')
        
        def save_edits():
            # Collect data
            new_data = {}
            new_data['Date'] = ent_date.get_date().strftime('%Y-%m-%d')
            new_data['Site'] = cb_site.get().strip()
            
            full_mat_name = cb_mat.get().strip()
            # Find MaterialID
            found_id = None
            for _, m_row in self.materials_df.iterrows():
                m_name = str(m_row.get('품목명', ''))
                m_model = str(m_row.get('모델명', '')).replace('nan', '').replace('None', '')
                m_sn = str(m_row.get('SN', '')).replace('nan', '').replace('None', '')
                m_display = m_name
                if m_model: m_display += f" - {m_model}"
                if m_sn: m_display += f" (SN: {m_sn})"
                if m_display == full_mat_name:
                    found_id = m_row['MaterialID']
                    break
            new_data['MaterialID'] = found_id if found_id is not None else entry_data.get('MaterialID')
            
            new_data['장비명'] = cb_equip.get().strip()
            new_data['검사방법'] = cb_method.get().strip()
            
            # Numeric fields
            for key in ['검사량', 'FilmCount', '단가', '출장비', '일식', '검사비']:
                try: 
                    v_str = fields[key].get().strip().replace(',', '')
                    new_data[key] = float(v_str) if v_str else 0.0
                except: new_data[key] = 0.0
                
            for cat in rtk_cats:
                try: new_data[f'RTK_{cat}'] = float(fields[f'RTK_{cat}'].get())
                except: new_data[f'RTK_{cat}'] = 0.0
                
            for cat in ndt_cats:
                m_key = "".join(cat.split())
                try: new_data[f'NDT_{m_key}'] = float(fields[f'NDT_{m_key}'].get())
                except: new_data[f'NDT_{m_key}'] = 0.0
                
            for key in ['차량번호', '주행거리', '차량점검', '차량비고']:
                new_data[key] = fields[key].get().strip()
                
            for i in range(1, 11):
                name_key = 'User' if i == 1 else f'User{i}'
                time_key = 'WorkTime' if i == 1 else f'WorkTime{i}'
                ot_key = 'OT' if i == 1 else f'OT{i}'
                
                new_data[name_key] = worker_fields[f'name{i}'].get().strip()
                new_data[time_key] = worker_fields[f'time{i}'].get().strip()
                new_data[ot_key] = worker_fields[f'ot{i}'].get().strip()

            # Provide immediate feedback
            btn_save_edit.config(state='disabled', text="저장 중...")
            edit_win.update_idletasks()
            
            try:
                if self.save_daily_usage_edits(df_idx, new_data):
                    edit_win.destroy()
                    messagebox.showinfo("수정 완료", "기록이 성공적으로 수정되었으며 재고 내역이 업데이트되었습니다.")
                else:
                    btn_save_edit.config(state='normal', text="주진철 수정완료")
            except Exception as e:
                btn_save_edit.config(state='normal', text="주진철 수정완료")
                messagebox.showerror("오류", f"수정 중 오류가 발생했습니다: {e}")

        btn_save_edit = ttk.Button(btn_frame, text="주진철 수정완료", style='Big.TButton', command=save_edits)
        btn_save_edit.pack(side='right', padx=10)
        ttk.Button(btn_frame, text="취소", command=edit_win.destroy).pack(side='right', padx=10)

    def _calculate_ot_from_worktime(self, worktime_value, calculation_date):
        """Standalone helper to calculate OT hours and amount from a worktime string"""
        try:
            if not worktime_value:
                return 0.0, 0
                
            import re
            marker_pattern = MARKER_PATTERN
            clean_val = marker_pattern.sub('', str(worktime_value)).strip()
            
            # Use regex to find separator (~ or -) more robustly
            sep_match = re.search(r'[:\d]\s*([~-])\s*[:\d]', clean_val)
            if not sep_match: return 0.0, 0
            sep = sep_match.group(1)
            
            start_time_str, end_time_str = clean_val.split(sep)
            sh, sm = map(int, start_time_str.split(':'))
            eh, em = map(int, end_time_str.split(':'))
            
            start_f = sh + sm / 60.0
            end_f = eh + em / 60.0
            if end_f < start_f: end_f += 24
            total_duration = end_f - start_f
            if total_duration <= 0: return 0.0, 0

            weekday = calculation_date.weekday()
            is_holiday = weekday >= 5
            is_friday = (weekday == 4)

            ot_hours = 0.0
            amount = 0
            if is_holiday:
                ot_hours = total_duration
                amount = ot_hours * 7500
            else:
                if end_f > 18:
                    ot_start = max(start_f, 18.0)
                    ot_hours = end_f - ot_start
                    evening_end = min(end_f, 22.0)
                    evening_hours = max(0, evening_end - ot_start)
                    night_start = max(ot_start, 22.0)
                    night_end = min(end_f, 24.0)
                    night_hours = max(0, night_end - night_start)
                    dawn_start = max(ot_start, 24.0)
                    dawn_hours = max(0, end_f - dawn_start)
                    dawn_rate = 7500 if is_friday else 5000
                    amount = (evening_hours * 4000) + (night_hours * 5000) + (dawn_hours * dawn_rate)
            
            return ot_hours, int(amount)
        except:
            return 0.0, 0

    def calculate_and_update_ot_manual(self, worktime_value, ot_entry, calculation_date=None):
        """Helper for personal calculation during editing without complex widget lookups"""
        try:
            current_date = calculation_date if calculation_date else self.ent_daily_date.get_date()
            ot_hours, amount = self._calculate_ot_from_worktime(worktime_value, current_date)
            
            if amount > 0:
                ot_entry.delete(0, tk.END)
                ot_entry.insert(0, f"{amount:,}")
            elif ot_hours == 0:
                # If no OT, clear it or set to 0
                ot_entry.delete(0, tk.END)
                ot_entry.insert(0, "0")
        except: pass

    def save_daily_usage_edits(self, df_idx, new_data):
        """Save edited daily usage and reconcile stock"""
        try:
            old_entry = self.daily_usage_df.loc[df_idx]
            
            # 1. Revert Old Stock (Delete transactions)
            old_site = old_entry.get('Site', '')
            old_date = pd.to_datetime(old_entry.get('Date'))
            old_mat_id = old_entry.get('MaterialID')
            note_pattern = f'{old_site} 현장 사용 (자동 차감)'
            
            if not self.transactions_df.empty:
                trans_mask = (
                    (pd.to_datetime(self.transactions_df['Date']).dt.date == old_date.date()) &
                    (self.transactions_df['MaterialID'] == old_mat_id) &
                    (self.transactions_df['Type'] == 'OUT') &
                    (self.transactions_df['Note'] == note_pattern)
                )
                
                # Revert NDT too
                ndt_names = ["형광자분", "흑색자분", "백색페인트", "침투제", "세척제", "현상제", "형광침투제"]
                ndt_ids = []
                for cat in ndt_names:
                    m_key = "".join(cat.split())
                    if old_entry.get(f'NDT_{m_key}', 0) > 0:
                        m_rows = self.materials_df[self.materials_df['품목명'] == cat]
                        if not m_rows.empty:
                            ndt_ids.append(m_rows.iloc[0]['MaterialID'])
                
                if ndt_ids:
                    trans_mask |= (
                        (pd.to_datetime(self.transactions_df['Date']).dt.date == old_date.date()) &
                        (self.transactions_df['MaterialID'].isin(ndt_ids)) &
                        (self.transactions_df['Type'] == 'OUT') &
                        (self.transactions_df['Note'] == note_pattern)
                    )
                
                self.transactions_df = self.transactions_df[~trans_mask]

            # 2. Update entry in DF
            for k, v in new_data.items():
                if k == 'MaterialID':
                    try: self.daily_usage_df.at[df_idx, k] = float(v)
                    except: pass
                else:
                    self.daily_usage_df.at[df_idx, k] = v
            
            # Re-apply calculated total fields
            rtk_total = sum([new_data.get(f'RTK_{c}', 0) for c in ["센터미스", "농도", "마킹미스", "필름마크", "취급부주의", "고객불만", "기타"]])
            self.daily_usage_df.at[df_idx, 'Usage'] = rtk_total
            
            # 3. Apply New Deduction
            new_date = pd.to_datetime(new_data['Date'])
            new_site = new_data['Site']
            new_mat_id = new_data['MaterialID']
            new_film_count = new_data.get('FilmCount', 0)
            new_note_pattern = f'{new_site} 현장 사용 (자동 차감)'
            
            # Collect all workers names for transaction
            workers_names = []
            for i in range(1, 11):
                n = new_data.get('User' if i == 1 else f'User{i}', '').strip()
                if n: workers_names.append(n)
            all_workers = ", ".join(workers_names)
            
            # Use same logic as add_daily_usage_entry but adapted
            total_deduction = new_film_count + rtk_total
            if total_deduction > 0:
                new_trans = {
                    'Date': datetime.datetime.combine(new_date.date(), datetime.datetime.now().time()),
                    'MaterialID': new_mat_id,
                    'Type': 'OUT',
                    'Quantity': total_deduction,
                    'User': all_workers,
                    'Site': new_site,
                    'Note': new_note_pattern
                }
                self.transactions_df = pd.concat([self.transactions_df, pd.DataFrame([new_trans])], ignore_index=True)
            
            # NDT Deductions
            for cat in ndt_names:
                m_key = "".join(cat.split())
                qty = new_data.get(f'NDT_{m_key}', 0)
                if qty > 0:
                    found_ids = self.materials_df[self.materials_df['품목명'] == cat]['MaterialID']
                    if not found_ids.empty:
                        ndt_trans = {
                            'Date': datetime.datetime.combine(new_date.date(), datetime.datetime.now().time()),
                            'MaterialID': found_ids.iloc[0],
                            'Type': 'OUT',
                            'Quantity': qty,
                            'User': all_workers,
                            'Site': new_site,
                            'Note': new_note_pattern
                        }
                        self.transactions_df = pd.concat([self.transactions_df, pd.DataFrame([ndt_trans])], ignore_index=True)

            # 4. Save & Refresh
            if self.save_data():
                self.update_daily_usage_view()
                self.update_stock_view()
                self.update_transaction_view()
                self.update_material_combo() # Refresh history suggestions
                return True
            return False
        except Exception as e:
            messagebox.showerror("수정 오류", f"기록 수정 중 오류가 발생했습니다: {e}")
            return False
                
    def export_daily_usage_history(self):
        """Export current filtered daily usage history to Excel"""
        try:
            # Get selected items only, or all if none selected
            selected_items = self.daily_usage_tree.selection()
            if not selected_items:
                # If no items selected, export all visible items
                items_to_export = self.daily_usage_tree.get_children()
            else:
                # Export only selected items
                items_to_export = selected_items
            
            data = []
            for item in items_to_export:
                data.append(self.daily_usage_tree.item(item, 'values'))
            
            if not data:
                messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
                return
            
            # Export columns currently visible in the Treeview
            display_cols = self.daily_usage_tree['displaycolumns']
            if not display_cols or display_cols == ('#all'):
                selected_cols = list(self.daily_usage_tree['columns'])
            else:
                selected_cols = list(display_cols)
            
            # Map selected column names to their indices in the Treeview values
            all_cols = list(self.daily_usage_tree['columns'])
            col_indices = [all_cols.index(col) for col in selected_cols]
            
            # Reconstruct data
            filtered_data = []
            for item_values in data:
                row = [item_values[i] for i in col_indices if i < len(item_values)]
                filtered_data.append(row)
            
            filename = f"일일사용내역_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=filename, filetypes=[("Excel files", "*.xlsx")])
            
            if save_path:
                cols = list(self.daily_usage_tree['columns'])
                col_to_idx = {col: i for i, col in enumerate(cols)}
                
                # Rebuild data from 'data' (which has ALL columns including hidden ones)
                # and swap summarized '작업자' with full list from '(Full작업자)'
                final_filtered_data = []
                for item_values in data:
                    row_list = list(item_values)
                    
                    # Perform the swap if both columns exist
                    if '작업자' in cols and '(Full작업자)' in col_to_idx:
                        full_idx = col_to_idx['(Full작업자)']
                        if full_idx < len(row_list):
                            full_val = row_list[full_idx]
                            if full_val:
                                # Find index of '작업자' in ALL columns to swap in row_list
                                idx_worker = cols.index('작업자')
                                row_list[idx_worker] = full_val
                    
                    # Extract only the selected display columns for final export
                    final_row = [row_list[i] for i in col_indices if i < len(row_list)]
                    final_filtered_data.append(final_row)
                
                df = pd.DataFrame(final_filtered_data, columns=selected_cols)
                df = self.clean_df_export(df)
                self.save_df_to_excel_autofit(df, save_path, "일일사용내역")
                messagebox.showinfo("완료", f"데이터가 엑셀로 저장되었습니다.\n{save_path}")
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 내보내기 중 오류가 발생했습니다: {e}")

    def show_column_visibility_dialog(self):
        """Open dialog to manually show/hide columns in the history view"""
        all_cols = list(self.daily_usage_tree['columns'])
        
        # Mandatory columns are usually kept, but we let user choose most of them
        # only excluding internal technical columns
        selectable_cols = [c for c in all_cols if c != '(Full작업자)']
        
        # Pre-select columns that are currently visible
        active_cols = self.daily_usage_tree['displaycolumns']
        if not active_cols or active_cols == ('#all'):
             active_cols = all_cols
        
        dialog = ColumnSelectionDialog(self.root, selectable_cols, title="표시 컬럼 관리")
        # Overwrite vars with current visibility state
        for col, var in dialog.vars.items():
            var.set(col in active_cols)
            
        dialog.wait_window()
        
        if dialog.result is not None:
            # Result set from dialog
            self.manual_visible_cols = dialog.result
            
            # Reconstruct final display based on user result + always mandatory ones
            mandatory_cols = ['날짜', '현장', '작업자', '품목명', '입력시간']
            final_selection = list(dialog.result)
            for mc in mandatory_cols:
                if mc in all_cols and mc not in final_selection:
                    final_selection.append(mc)
            
            # Re-order to match original tree columns
            sorted_selection = [c for c in all_cols if c in final_selection]
            
            self.daily_usage_tree['displaycolumns'] = sorted_selection
            self.save_tab_config()
            
            # Immediate refresh to apply selection
            self.update_daily_usage_view()

    def export_all_daily_usage(self):
        """Export all daily usage records to Excel"""
        try:
            if self.daily_usage_df.empty:
                messagebox.showinfo("알림", "기록된 데이터가 없습니다.")
                return
            
            # Export all columns from the tree structure
            selected_cols = list(self.daily_usage_tree['columns'])
            
            filename = f"전체사용기록_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=filename, filetypes=[("Excel files", "*.xlsx")])
            
            if save_path:
                # Ensure we only try to export columns that exist in the dataframe
                valid_cols = [c for c in selected_cols if c in self.daily_usage_df.columns]
                # If dataframe has extra columns not in tree (like raw data), include them? 
                # Better to stick to what the user sees/defines in tree interface + critical data
                # Actually, export_all usually implies backing up the raw dataframe. 
                # But previous logic used tree columns. Let's stick to tree columns if they exist in DF.
                
                # Fallback: if tree columns are not in DF (e.g. calculated columns), we might have issues.
                # The daily_usage_df has raw data. The tree has formatted data.
                # daily_usage_df columns: Date, Site, MaterialID, Usage, Note, etc. + RTK_..., NDT_...
                # Tree columns: 날짜, 현장, 작업자... OT합계...
                
                # If we want "All Daily Usage", we should probably export the RAW dataframe for backup purposes, 
                # OR the formatted view for reporting.
                # The previous code seemed to try to export based on tree columns but mapped from DF?
                # "export_df = self.daily_usage_df[selected_cols].copy()"
                # This implies 'selected_cols' MUST exist in daily_usage_df.
                # BUT 'daily_usage_tree['columns']' includes 'OT합계', 'RTK총계' which are likely NOT in daily_usage_df (calculated).
                
                # Let's check daily_usage_df columns again.
                # It has 'Date', 'Site', 'MaterialID'...
                # It does NOT have '날짜', '현장' (Korean names).
                # So the previous code 'self.daily_usage_df[selected_cols]' would have FAILED if selected_cols came from tree columns!
                
                # WAIT. The user said "remove column selection". 
                # The previous working code (before my changes) probably did:
                # export_df = self.daily_usage_df.copy()
                # Let's revert to a safe "Backup" style export for "All Data".
                
                export_df = self.daily_usage_df.copy()
                export_df = self.clean_df_export(export_df)
                self.save_df_to_excel_autofit(export_df, save_path, "전체사용기록")
                
                export_df = self.clean_df_export(export_df)
                self.save_df_to_excel_autofit(export_df, save_path, "전체기록")
                messagebox.showinfo("완료", f"전체 기록이 엑셀로 저장되었습니다.\n{save_path}")
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 내보내기 중 오류가 발생했습니다: {e}")


    def setup_keyboard_shortcuts(self):
        """Setup keyboard shortcuts for navigation"""
        # Ctrl+Tab to switch between notebook tabs (forward)
        self.root.bind('<Control-Tab>', self.next_tab)
        # Ctrl+Shift+Tab to switch between notebook tabs (backward)
        self.root.bind('<Control-Shift-Tab>', self.prev_tab)
        
        # Alt+숫자 for direct tab access
        self.root.bind('<Alt-Key-1>', lambda e: self.notebook.select(0))
        self.root.bind('<Alt-Key-2>', lambda e: self.notebook.select(1))
        self.root.bind('<Alt-Key-3>', lambda e: self.notebook.select(2))
        self.root.bind('<Alt-Key-4>', lambda e: self.notebook.select(3))
        self.root.bind('<Alt-Key-5>', lambda e: self.notebook.select(4))
        self.root.bind('<Alt-Key-6>', lambda e: self.notebook.select(5))
        
        # Right-click on notebook for tab reordering
        # Tab interactions removed here, kept in create_widgets
    
    def next_tab(self, event=None):
        """Switch to next tab"""
        current = self.notebook.index(self.notebook.select())
        total = self.notebook.index('end')
        next_tab = (current + 1) % total
        self.notebook.select(next_tab)
        return 'break'
    
    def prev_tab(self, event=None):
        """Switch to previous tab"""
        current = self.notebook.index(self.notebook.select())
        total = self.notebook.index('end')
        prev_tab = (current - 1) % total
        self.notebook.select(prev_tab)
        return 'break'
    
    def show_tab_context_menu(self, event):
        """Show context menu for tab reordering"""
        # Identify which tab was clicked
        try:
            clicked_tab = self.notebook.index(f"@{event.x},{event.y}")
            self.notebook.select(clicked_tab)
            
            # Create context menu
            context_menu = tk.Menu(self.root, tearoff=0)
            
            # Only show "Move Left" if not the first tab
            if clicked_tab > 0:
                context_menu.add_command(label="← 탭 왼쪽으로 이동", 
                                        command=lambda: self.move_tab_left(clicked_tab))
            
            # Only show "Move Right" if not the last tab
            if clicked_tab < self.notebook.index('end') - 1:
                context_menu.add_command(label="탭 오른쪽으로 이동 →", 
                                        command=lambda: self.move_tab_right(clicked_tab))
            
            # Show menu at cursor position
            context_menu.post(event.x_root, event.y_root)
        except:
            pass
    
    def move_tab_left(self, tab_index):
        """Move tab one position to the left"""
        if tab_index > 0:
            # Get tab info
            tab = self.notebook.tabs()[tab_index]
            text = self.notebook.tab(tab_index, "text")
            
            # Remove and reinsert at new position
            self.notebook.insert(tab_index - 1, tab, text=text)
            self.notebook.select(tab_index - 1)
            
            # Save configuration immediately
            self.save_tab_config()
    
    def move_tab_right(self, tab_index):
        """Move tab one position to the right"""
        if tab_index < self.notebook.index('end') - 1:
            # Get tab info
            tab = self.notebook.tabs()[tab_index]
            text = self.notebook.tab(tab_index, "text")
            
            # Remove and reinsert at new position
            self.notebook.insert(tab_index + 2, tab, text=text)
            self.notebook.select(tab_index + 1)
            
            # Save configuration immediately
            self.save_tab_config()

    def on_tab_drag_start(self, event):
        """Start tab dragging by identifying the tab under the cursor"""
        try:
            # Check if click is on a tab tab
            clicked_tab = self.notebook.identify(event.x, event.y)
            if clicked_tab == "label":
                # Find which index this is
                self._drag_start_index = self.notebook.index(f"@{event.x},{event.y}")
                self._current_drag_index = self._drag_start_index
            else:
                self._drag_start_index = None
        except:
            self._drag_start_index = None

    def on_tab_drag(self, event):
        """Handle visual swapping of tabs during drag"""
        if not hasattr(self, '_drag_start_index') or self._drag_start_index is None:
            return

        try:
            # Find the index of the tab currently under the cursor
            target_index = self.notebook.index(f"@{event.x},{event.y}")
            
            if target_index != self._current_drag_index:
                # Get the widget of the tab we are dragging
                tab_widget = self.notebook.tabs()[self._current_drag_index]
                tab_text = self.notebook.tab(self._current_drag_index, "text")
                
                # Use insert to move tab widget. 
                # Note: insert(pos, widget) handles the reordering logic in Notebook
                self.notebook.insert(target_index, tab_widget, text=tab_text)
                
                self._current_drag_index = target_index
                self.notebook.select(target_index)
        except:
            pass

    def on_tab_drag_end(self, event):
        """Finalize tab order and save configuration"""
        if hasattr(self, '_drag_start_index') and self._drag_start_index is not None:
            if self._current_drag_index != self._drag_start_index:
                print(f"Tab reordered: {self._drag_start_index} -> {self._current_drag_index}")
                self.save_tab_config()
            
            self._drag_start_index = None
            self._current_drag_index = None
    
    def on_tab_changed(self, event=None):
        """Handle tab selection change event"""
        try:
            # 1. Save configuration when tab changes (respects is_ready via save_tab_config)
            self.save_tab_config()

            # 1-1. 탭 이동 시 실제 데이터(Excel)도 자동 저장
            if getattr(self, 'is_ready', False):
                self.save_data()
            
            # 2. Handle specific tab UI adjustments
            current_tab = self.notebook.select()
            if not current_tab:
                return
                
            # Convert widget path to index if needed
            try:
                current_tab_idx = self.notebook.index("current")
                tab_text = self.notebook.tab(current_tab_idx, "text")
            except:
                tab_text = ""

            # Check for Daily Usage tab
            if (hasattr(self, 'tab_daily_usage') and str(current_tab) == str(self.tab_daily_usage)) or \
               tab_text == '현장별 일일 사용량 기입':
                
                print("Daily usage tab selected - ensuring visibility")
                # Force multiple updates when tab is selected
                self.root.after(50, self._ensure_daily_usage_sash_visibility)
                self.root.after(200, self._ensure_daily_usage_sash_visibility)
                self.root.after(400, self._ensure_canvas_scroll_region)
                
                # Also ensure the inner history sash is visible
                self.root.after(150, self._ensure_sash_visible)
            
            elif tab_text == '월별 집계':
                print("Monthly usage tab selected - refreshing view")
                self.update_monthly_usage_view()
        except Exception as e:
            print(f"Error in tab change handler: {e}")
    
    def auto_save_to_list(self, event, combobox, data_list, config_key):
        """Helper to auto-save new entry from combobox to a list and update all related UI"""
        new_val = combobox.get().strip()
        if not new_val:
            return
            
        if new_val not in data_list:
            data_list.append(new_val)
            data_list.sort()
            
            # Trigger app-wide update
            self.refresh_ui_for_list_change(config_key)
            self.save_tab_config()

    def auto_save_worktime(self, event, entry, config_key):
        """Helper to auto-save worktime values and support copy functionality"""
        worktime_value = entry.get().strip()
        if not worktime_value:
            return
            
        # Initialize worktimes list if not exists
        if not hasattr(self, 'worktimes'):
            self.worktimes = []
            
        # Check if this value already exists
        if worktime_value not in self.worktimes:
            self.worktimes.append(worktime_value)
            self.worktimes.sort()
            
            # Save to config and update UI
            try:
                self.refresh_ui_for_list_change('worktimes')
                self.save_tab_config()
            except Exception as e:
                print(f"Failed to save worktime config: {e}")
        
        # Calculate and update corresponding OT field
        self.calculate_and_update_ot(worktime_value, entry)
        
        # Add copy functionality - select all text on focus and bind Ctrl+C (only once)
        entry.select_range(0, tk.END)
        if not hasattr(entry, '_copy_bound'):
            entry.bind('<Control-c>', lambda e: self.copy_worktime(entry))
            entry._copy_bound = True
        
        # Store current value for copy functionality
        entry.last_value = worktime_value

    def calculate_and_update_ot(self, worktime_value, worktime_entry):
        """Calculate OT amount based on worktime and update corresponding OT field"""
        try:
            current_date = self.ent_daily_date.get_date()
            ot_hours, amount = self._calculate_ot_from_worktime(worktime_value, current_date)
            
            ot_entry = self.get_corresponding_ot_field(worktime_entry)
            if ot_entry:
                if amount > 0:
                    ot_entry.delete(0, tk.END)
                    ot_entry.insert(0, f"{amount:,}")
                else:
                    ot_entry.delete(0, tk.END)
                    # If it was cleared, but hours > 0 (very small amount?), maybe set to 0
                    if ot_hours > 0: ot_entry.insert(0, "0")
        except Exception as e:
            print(f"Error calculating OT: {e}")

    def get_corresponding_ot_field(self, worktime_entry):
        """Get corresponding OT field for a worktime entry"""
        # Map worktime entries to OT entries
        worktime_to_ot = {
            'ent_worktime1': 'ent_ot1', 'ent_worktime2': 'ent_ot2', 
            'ent_worktime3': 'ent_ot3', 'ent_worktime4': 'ent_ot4',
            'ent_worktime5': 'ent_ot5', 'ent_worktime6': 'ent_ot6',
            'ent_worktime7': 'ent_ot7', 'ent_worktime8': 'ent_ot8',
            'ent_worktime9': 'ent_ot9', 'ent_worktime10': 'ent_ot10'
        }
            
        # Find worktime entry name by checking object IDs
        worktime_id = id(worktime_entry)
            
        worktime_attrs = [f'ent_worktime{i}' for i in range(1, 11)]
        for attr_name in worktime_attrs:
            if hasattr(self, attr_name):
                attr_value = getattr(self, attr_name)
                if id(attr_value) == worktime_id:
                    ot_attr_name = worktime_to_ot.get(attr_name)
                    if ot_attr_name and hasattr(self, ot_attr_name):
                        return getattr(self, ot_attr_name)
                    break
            
        return None

    def copy_worktime(self, entry):
        """Copy worktime value to clipboard"""
        try:
            worktime_value = entry.get().strip()
            if worktime_value:
                self.root.clipboard_clear()
                self.root.clipboard_append(worktime_value)
                # Show brief feedback
                entry.delete(0, tk.END)
                entry.insert(0, worktime_value)
        except Exception as e:
            pass  # Silently handle clipboard errors

    def auto_save_ot(self, event, entry, config_key):
        """Helper to auto-save OT values and support copy functionality"""
        ot_value = entry.get().strip()
        if not ot_value:
            return
                
        # Initialize ot_times list if not exists
        if not hasattr(self, 'ot_times'):
            self.ot_times = []
            
        # Calculate OT amount for any input
        calculated_amount = self.calculate_ot_amount(ot_value)
        if calculated_amount:
            # Update entry with calculated amount
            entry.delete(0, tk.END)
            display_value = f"{ot_value} ({calculated_amount:,}원)"
            entry.insert(0, display_value)
            ot_value = display_value
            print(f"OT calculated: {ot_value}")  # Debug print
            
        # Check if this value already exists
        if ot_value not in self.ot_times:
            self.ot_times.append(ot_value)
            self.ot_times.sort()
            
        # Save to config with error handling - only save when list changes
        try:
            self.save_tab_config()
        except Exception as e:
            print(f"Failed to save OT config: {e}")
            
        # Add copy functionality - select all text on focus and bind Ctrl+C (only once)
        entry.select_range(0, tk.END)
        if not hasattr(entry, '_copy_bound'):
            entry.bind('<Control-c>', lambda e: self.copy_ot(entry))
            entry._copy_bound = True
            
        # Store current value for copy functionality
        entry.last_value = ot_value

    def calculate_ot_amount(self, ot_value):
        """Calculate OT amount based on time and rates, or just parse if already an amount"""
        try:
            if not ot_value or not str(ot_value).strip(): return 0
            val = str(ot_value).strip().replace(',', '')
            
            # If it's already just a large number, it's the amount
            if val.isdigit() and int(val) > 100:
                return int(val)
            
            # If it has (N원) format
            if '(' in val and '원)' in val:
                try:
                    return int(val.split('(')[1].split('원')[0].replace(',', ''))
                except: pass

            hours = self._parse_ot_hours(val)
            if hours <= 0: return 0

            # (Rest of simple duration fallback)
            evening_hours = min(hours, 4)
            night_hours = max(0, hours - 4)
            return int(evening_hours * 4000 + night_hours * 5000)
        except Exception as e:
            return 0

    def _parse_ot_hours(self, ot_value):
        """Helper to extract numeric OT hours using regex for maximum robustness"""
        import re
        try:
            if not ot_value or not str(ot_value).strip(): return 0
            val = str(ot_value).strip().replace(' ', '').replace('익일', '')
            
            # If it's just a large number (>100), assume it's an amount, not hours
            if val.replace(',', '').isdigit() and int(val.replace(',', '')) > 100:
                return 0.0

            # 1. Check for "N시간"
            dur_match = re.search(r'(\d+\.?\d*)\s*(시간|hr|h)', val)
            if dur_match:
                return float(dur_match.group(1))

            # 2. Check for time range "18:00~22:00"
            range_match = re.search(r'(\d{1,2}):(\d{1,2})[-~](\d{1,2}):(\d{1,2})', val)
            if range_match:
                h1, m1, h2, m2 = map(int, range_match.groups())
                if h2 < h1: h2 += 24
                return (h2 * 60 + m2 - (h1 * 60 + m1)) / 60

            # 3. Check for simple ':' format "2:30"
            colon_match = re.search(r'^(\d{1,2}):(\d{1,2})$', val)
            if colon_match:
                h, m = map(int, colon_match.groups())
                return h + (m / 60)

            # 4. Fallback to just extracting the first small float/int found
            num_match = re.search(r'(\d+\.?\d*)', val)
            if num_match:
                v = float(num_match.group(1))
                if v <= 24: return v # Reasonable hour count
            
            return 0
        except:
            return 0

    def _calculate_split_ot_hours(self, ot_value, date_val=None):
        """Split OT hours into Day window (18-22), Night window (22-24), and Holiday window for weekends/Friday dawn"""
        import re
        import pandas as pd
        try:
            if not ot_value or not str(ot_value).strip(): return 0.0, 0.0, 0.0
            val = str(ot_value).strip().replace(' ', '')
            
            total_hours = self._parse_ot_hours(val)
            if total_hours <= 0: return 0.0, 0.0, 0.0

            # Default start at 18:00 if no range
            start_hour = 18
            range_match = re.search(r'(\d{1,2}):(\d{1,2})[-~](\d{1,2}):(\d{1,2})', val)
            if range_match:
                start_hour = int(range_match.group(1))
            
            # Handle overnight logic (e.g., 18:00~01:00)
            current_time = float(start_hour)
            remaining = total_hours
            day_hours = 0.0
            night_hours = 0.0
            holiday_hours = 0.0
            
            # Check if this is Friday going into Saturday
            is_friday = False
            if date_val is not None:
                try:
                    dt = pd.to_datetime(date_val)
                    is_friday = (dt.weekday() == 4)
                except:
                    pass
            
            # Simulate hour by hour (or portion by portion)
            while remaining > 0:
                # 18:00 ~ 22:00 구간은 연장근무 (day_hours)
                if 18 <= current_time < 22:
                    can_take = 22 - current_time
                    taken = min(remaining, can_take)
                    day_hours += taken
                    remaining -= taken
                    current_time += taken
                # 22:00 ~ 24:00 구간은 야간근무 (night_hours)
                elif 22 <= current_time < 24:
                    can_take = 24 - current_time
                    taken = min(remaining, can_take)
                    night_hours += taken
                    remaining -= taken
                    current_time += taken
                # 24:00 ~ (익일) 구간
                elif current_time >= 24:
                    can_take = remaining # 끝까지 처리
                    taken = min(remaining, can_take)
                    if is_friday:
                        holiday_hours += taken
                    else:
                        night_hours += taken
                    remaining -= taken
                    current_time += taken
                # 18:00 이전 시간이 혹시 OT로 입력되었다면 상황에 맞게 처리 (기본 시뮬레이션에서는 18시까지 대기하는 것으로 가정)
                else: 
                     current_time = 18.0
            
            return day_hours, night_hours, holiday_hours
        except:
            return 0.0, 0.0, 0.0

    def copy_ot(self, entry):
        """Copy OT value to clipboard"""
        try:
            ot_value = entry.get().strip()
            if ot_value:
                self.root.clipboard_clear()
                self.root.clipboard_append(ot_value)
                # Show brief feedback
                entry.delete(0, tk.END)
                entry.insert(0, ot_value)
        except Exception as e:
            pass  # Silently handle clipboard errors


    def save_tab_config(self, force=False):
        """Save current tab configuration to memory and disk"""
        try:
            # [STABILITY] Skip saving if app is not fully ready (prevents overwriting saved tab index during init)
            if not getattr(self, 'is_ready', False) and not force:
                return

            # Force update to ensure all coordinates and sizes are accurate
            self.root.update_idletasks()
            
            # Initialize or keep existing tab_config
            if not hasattr(self, 'tab_config'):
                self.tab_config = {}
                
            # Update with current core values
            current_tab_idx = 0
            try:
                current_tab_idx = self.notebook.index(self.notebook.select())
            except:
                current_tab_idx = self.tab_config.get('selected_tab', 0)

            self.tab_config.update({
                'selected_tab': current_tab_idx,
                'tab_order': [],
                'sites': self.sites,
                'users': getattr(self, 'users', []),
                'warehouses': getattr(self, 'warehouses', []),
                'equipments': getattr(self, 'equipments', []),
                'vehicles': getattr(self, 'vehicles', []),
                'worktimes': getattr(self, 'worktimes', []),
                'ot_times': getattr(self, 'ot_times', []),
                'layout_locked': self.layout_locked,
                'resolution_locked': self.resolution_locked,
                'locked_width': getattr(self, 'locked_width', 1200),
                'locked_height': getattr(self, 'locked_height', 800),
                'daily_usage_sash_locked': self.daily_usage_sash_locked,
                'daily_usage_sash_pos': self.daily_usage_paned.sashpos(0) if hasattr(self, 'daily_usage_paned') else None,
                'daily_history_sash_pos': self.daily_history_paned.sashpos(0) if hasattr(self, 'daily_history_paned') else None,
                'entry_inner_frame_height': self.entry_inner_frame.winfo_height() if hasattr(self, 'entry_inner_frame') else None,
                'history_visible_cols': getattr(self, 'manual_visible_cols', []),
                'monthly_visible_cols': getattr(self, 'monthly_visible_cols', []),
                'window_state': self.root.state(),
                'window_width': self.root.winfo_width(),
                'window_height': self.root.winfo_height()
            })
            
            # Save current stock column widths
            self.tab_config['stock_col_widths'] = {}
            if hasattr(self, 'stock_tree'):
                for col in self.stock_tree['columns']:
                    self.tab_config['stock_col_widths'][col] = self.stock_tree.column(col, 'width')

            # Save daily usage column widths
            self.tab_config['daily_usage_col_widths'] = {}
            if hasattr(self, 'daily_usage_tree'):
                for col in self.daily_usage_tree['columns']:
                    self.tab_config['daily_usage_col_widths'][col] = self.daily_usage_tree.column(col, 'width')
            
            # Save monthly usage column widths
            self.tab_config['monthly_usage_col_widths'] = {}
            if hasattr(self, 'monthly_usage_tree'):
                for col in self.monthly_usage_tree['columns']:
                    self.tab_config['monthly_usage_col_widths'][col] = self.monthly_usage_tree.column(col, 'width')
            
            # [NEW] Save summary table column widths
            self.tab_config['site_summary_col_widths'] = {}
            if hasattr(self, 'site_summary_tree'):
                for col in self.site_summary_tree['columns']:
                    self.tab_config['site_summary_col_widths'][col] = self.site_summary_tree.column(col, 'width')
            
            self.tab_config['worker_summary_col_widths'] = {}
            if hasattr(self, 'worker_summary_tree'):
                for col in self.worker_summary_tree['columns']:
                    self.tab_config['worker_summary_col_widths'][col] = self.worker_summary_tree.column(col, 'width')
            
            # [NEW] Save report column widths
            self.tab_config['report_col_widths'] = {}
            if hasattr(self, 'report_tree'):
                for col in self.report_tree['columns']:
                    self.tab_config['report_col_widths'][col] = self.report_tree.column(col, 'width')
            
            # Save tab order
            self.tab_config['tab_order'] = []
            for tab in self.notebook.tabs():
                self.tab_config['tab_order'].append(self.notebook.tab(tab, "text"))
            
            # Save Draggable items
            # Prepare in-memory geometries for merging (capture current states)
            current_geometries = {}
            
            for key, widget in self.draggable_items.items():
                manager = widget.winfo_manager()
                # Capture everything that is visible and managed
                if manager in ['place', 'grid']:
                    current_geometries[key] = {
                        'x': widget.winfo_x(),
                        'y': widget.winfo_y(),
                        'width': widget.winfo_width(),
                        'height': widget.winfo_height(),
                        'hidden': False
                    }
                    
                    if hasattr(widget, '_label_widget'):
                        current_geometries[key]['custom_label'] = widget._label_widget.cget('text')
                    if hasattr(widget, '_manage_list_key') and widget._manage_list_key:
                        current_geometries[key]['manage_list_key'] = widget._manage_list_key
                    if key.startswith('clone_'):
                        current_geometries[key]['is_clone'] = True
                        current_geometries[key]['widget_class_name'] = widget._widget_class.__name__
                        saved_kwargs = widget._widget_kwargs.copy()
                        if 'values' in saved_kwargs: del saved_kwargs['values']
                        current_geometries[key]['widget_kwargs'] = saved_kwargs
                    if key.startswith('memo_'):
                        current_geometries[key]['text'] = self.memos[key]['text_widget'].get('1.0', 'end-1c')
                        current_geometries[key]['memo_title'] = self.memos[key]['title_entry'].get()
                    if key in self.checklists:
                        current_geometries[key]['checklist_title'] = self.checklists[key]['title_entry'].get()
                        items_data = []
                        for child in self.checklists[key]['item_frame'].winfo_children():
                            if hasattr(child, '_checklist_data'):
                                items_data.append({
                                    'text': child._checklist_data['entry'].get(),
                                    'checked': child._checklist_data['var'].get()
                                })
                        current_geometries[key]['checklist_items'] = items_data
                    if key in self.vehicle_inspections:
                        current_geometries[key]['vehicle_data'] = self.vehicle_inspections[key].get_data()
                elif manager == '': # Hidden or not managed
                    # Only mark as hidden if it was previously in place/grid (to avoid junk)
                    if key in self.tab_config.get('draggable_geometries', {}):
                        current_geometries[key] = {'hidden': True}

            # [DEFINITIVE FIX] Deep merge current_geometries into disk config
            final_config = {}
            if os.path.exists(self.config_path):
                try:
                    with open(self.config_path, 'r', encoding='utf-8') as f:
                        final_config = json.load(f)
                except: pass
            
            # Use disk's geometries as base, then update with WHAT WE JUST CAPTURED in memory
            disk_geos = final_config.get('draggable_geometries', {})
            disk_geos.update(current_geometries)
            
            # Apply other tab state updates
            final_config.update(self.tab_config)
            final_config['draggable_geometries'] = disk_geos
            # [NEW] Save hidden_sites to config
            final_config['hidden_sites'] = list(getattr(self, 'hidden_sites', []))
            
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(final_config, f, ensure_ascii=False, indent=2)
            
            # Update in-memory copy as well
            self.tab_config = final_config
            
            print(f"Configuration saved. Locked: {self.tab_config.get('layout_locked')}, Tab: {self.tab_config.get('selected_tab')}")
        except Exception as e:
            print(f"Failed to save tab config: {e}")

    
    def load_tab_config(self):
        """Load and restore tab configuration"""
        try:
            if not hasattr(self, 'tab_config'):
                self.tab_config = {}
                
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    self.tab_config = json.load(f)
                    print(f"DEBUG: Loaded config raw: {self.tab_config}")
                
                config = self.tab_config
                
                # Apply high-level lock states immediately to internal variables
                self.layout_locked = config.get('layout_locked', False)
                self.resolution_locked = config.get('resolution_locked', False)
                self.daily_usage_sash_locked = config.get('daily_usage_sash_locked', False)
                self.locked_width = config.get('locked_width', 1200)
                self.locked_height = config.get('locked_height', 800)

                # Get current tab order
                current_order = []
                for tab in self.notebook.tabs():
                    tab_index = self.notebook.index(tab)
                    tab_text = self.notebook.tab(tab_index, "text")
                    current_order.append((tab_text, tab))
                
                # Create mapping from text to tab widget
                tab_map = {text: tab for text, tab in current_order}
                
                # Restore tab order if saved order exists
                saved_order = config.get('tab_order', [])
                if saved_order and len(saved_order) == len(current_order):
                    # Reorder tabs according to saved order
                    for i, tab_text in enumerate(saved_order):
                        if tab_text in tab_map:
                            tab = tab_map[tab_text]
                            current_pos = self.notebook.index(tab)
                            if current_pos != i:
                                self.notebook.insert(i, tab, text=tab_text)
                
                # Restore selected tab
                selected = config.get('selected_tab', 0)
                if 0 <= selected < self.notebook.index('end'):
                    try:
                        self.notebook.select(selected)
                    except: pass
                    # Force update after selection so the tab is rendered and children computed
                    self.root.update_idletasks()
                
                # Restore sites list - use in-place update
                self.sites[:] = config.get('sites', [])
                # If sites list is empty, try to populate from current daily_usage_df
                if not self.sites and not self.daily_usage_df.empty:
                    self.sites = sorted(self.daily_usage_df['Site'].dropna().unique().tolist())
                    self.sites = [str(s).strip() for s in self.sites if str(s).strip()]
                # [NEW] Restore hidden sites list
                self.hidden_sites[:] = config.get('hidden_sites', [])
                
                # Restore users list - use in-place update to preserve references
                self.users[:] = config.get('users', [])
                
                # [MIGRATION] Load and immediately migrate worktimes to ensure 익일 marker
                raw_worktimes = config.get('worktimes', [])
                self.worktimes = self._migrate_worktimes(raw_worktimes)
                
                # Restore vehicles and equipments
                self.vehicles[:] = config.get('vehicles', [])
                self.equipments[:] = config.get('equipments', [])

                
                # Refresh all WorkerDataGroup dropdowns with the migrated list
                for i in range(1, 11):
                    group_attr = f'worker_group{i}'
                    if hasattr(self, group_attr):
                        getattr(self, group_attr).update_time_list(self.worktimes)

                self.ot_times = config.get('ot_times', [])

                # [CLEANUP] Clean up users list from shift markers
                import re
                cleaned_users = []
                for user in self.users:
                    if re.match(r"^\((주간|야간|휴일)\)$", user.strip()):
                        continue
                    match = re.match(r"\((주간|야간|휴일)\)\s*(.*)", user.strip())
                    if match:
                        actual_name = match.group(2).strip()
                        if actual_name and actual_name not in cleaned_users:
                            cleaned_users.append(actual_name)
                    elif user.strip() and user.strip() not in cleaned_users:
                        cleaned_users.append(user.strip())
                
                if len(cleaned_users) != len(self.users):
                    self.users[:] = cleaned_users
                    # Don't save yet, let the normal flow handle it
                
                # Restore history visibility if saved
                self.manual_visible_cols = config.get('history_visible_cols', [])
                # [MIGRATION] If 'OT합계' is in saved cols, replace with 'OT시간', 'OT금액'
                if 'OT합계' in self.manual_visible_cols:
                    index = self.manual_visible_cols.index('OT합계')
                    self.manual_visible_cols.pop(index)
                    if 'OT시간' not in self.manual_visible_cols:
                        self.manual_visible_cols.insert(index, 'OT시간')
                    if 'OT금액' not in self.manual_visible_cols:
                        self.manual_visible_cols.insert(index + 1, 'OT금액')
                
                # [MIGRATION] Translate English names to Korean
                translations = {
                    'Date': '날짜', 'Site': '현장', 'User': '작업자', 'WorkTime': '작업시간',
                    'Equipment': '장비명', 'Method': '검사방법', 'Remark': '비고', 
                    'Note': '비고', 'EntryTime': '입력시간', 'MaterialName': '품목명'
                }
                for i, col in enumerate(self.manual_visible_cols):
                    if col in translations:
                        self.manual_visible_cols[i] = translations[col]
                
                if self.manual_visible_cols and hasattr(self, 'daily_usage_tree'):
                    try:
                        self.daily_usage_tree['displaycolumns'] = [c for c in self.manual_visible_cols if c in self.daily_usage_tree['columns']]
                    except: pass
                
                # Restore monthly history visibility if saved
                self.monthly_visible_cols = config.get('monthly_visible_cols', [])
                if self.monthly_visible_cols and hasattr(self, 'monthly_usage_tree'):
                    try:
                        self.monthly_usage_tree['displaycolumns'] = [c for c in self.monthly_visible_cols if c in self.monthly_usage_tree['columns']]
                    except: pass
                
                # Restore warehouses list
                self.warehouses[:] = config.get('warehouses', [])
                if not self.warehouses and not self.materials_df.empty:
                    self.warehouses = sorted(self.materials_df['창고'].dropna().unique().tolist())
                    self.warehouses = [str(w).strip() for w in self.warehouses if str(w).strip()]
                
                self.equipments[:] = config.get('equipments', [])
                if not self.equipments and not self.daily_usage_df.empty and '장비명' in self.daily_usage_df.columns:
                    self.equipments = sorted(self.daily_usage_df['장비명'].dropna().unique().tolist())
                    self.equipments = [str(e).strip() for e in self.equipments if str(e).strip()]
                
                # Vehicle list
                self.vehicles[:] = config.get('vehicles', [])
                
                # Restore stock column widths
                stock_col_widths = config.get('stock_col_widths', {})
                if stock_col_widths and hasattr(self, 'stock_tree'):
                    for col, width in stock_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.stock_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass
                
                # Restore daily usage column widths
                daily_usage_col_widths = config.get('daily_usage_col_widths', {})
                if daily_usage_col_widths and hasattr(self, 'daily_usage_tree'):
                    for col, width in daily_usage_col_widths.items():
                        try:
                            # Only apply if width is reasonable (e.g. > 10)
                            w = int(width)
                            
                            # All columns are now user-resizable
                            if w > 10:
                                # Enforce minimums for high-precision cols
                                if col == '날짜': w = max(w, 160)
                                elif col == '입력시간': w = max(w, 300)
                                self.daily_usage_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass

                # Restore monthly usage column widths
                monthly_usage_col_widths = config.get('monthly_usage_col_widths', {})
                if monthly_usage_col_widths and hasattr(self, 'monthly_usage_tree'):
                    for col, width in monthly_usage_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.monthly_usage_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass
                
                # [NEW] Restore site summary column widths
                site_summary_col_widths = config.get('site_summary_col_widths', {})
                if site_summary_col_widths and hasattr(self, 'site_summary_tree'):
                    for col, width in site_summary_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.site_summary_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass

                # [NEW] Restore worker summary column widths
                worker_summary_col_widths = config.get('worker_summary_col_widths', {})
                if worker_summary_col_widths and hasattr(self, 'worker_summary_tree'):
                    for col, width in worker_summary_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.worker_summary_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass
                # [NEW] Restore report column widths
                report_col_widths = config.get('report_col_widths', {})
                if report_col_widths and hasattr(self, 'report_tree'):
                    for col, width in report_col_widths.items():
                        try:
                            w = int(width)
                            if w > 10:
                                self.report_tree.column(col, width=w, minwidth=20, stretch=False)
                        except:
                            pass
                
                if hasattr(self, '_loading_memos'):
                    del self._loading_memos
                    
                # Restore layout lock state
                if hasattr(self, 'btn_lock_layout'):
                    if self.layout_locked:
                        self.btn_lock_layout.config(text="🔒 배치 고정됨")
                        self.style.configure("Lock.TButton", foreground="black")
                    else:
                        self.btn_lock_layout.config(text="🔓 배치 수정 중")
                        self.style.configure("Lock.TButton", foreground="red")

                if config.get('daily_usage_sash_locked', False):
                    self.daily_usage_sash_locked = True
                    
                    if hasattr(self, 'btn_sash_lock'):
                        self.btn_sash_lock.config(text="🔒 경계 잠금됨")
                    
                    print("LOADED: Restored sash lock state")

                # Tab selection handled already at line 4531
                
                # Recreate Memos and Clones first (these must exist before they can be placed)
                self._loading_memos = []
                # Map class names to actual classes for recreation
                class_map = {'Entry': ttk.Entry, 'Combobox': ttk.Combobox}
                
                draggable_geos = config.get('draggable_geometries', {})
                for key, geo in draggable_geos.items():
                    if key.startswith('memo_'):
                        self._loading_memos.append(key)
                        self.add_new_memo(
                            initial_text=geo.get('text', ""), 
                            initial_title=geo.get('memo_title', "메모"), 
                            key=key
                        )
                    elif key.startswith('checklist_'):
                        self._loading_memos.append(key)
                        self.add_new_checklist(
                            initial_data=geo.get('checklist_items', []),
                            initial_title=geo.get('checklist_title', "체크리스트"),
                            key=key
                        )
                    elif key.startswith('clone_'):
                        self._loading_memos.append(key)
                        cls_name = geo.get('widget_class_name', 'Entry')
                        cls = class_map.get(cls_name, ttk.Entry)
                        label = geo.get('custom_label', "복제항목")
                        kwargs = geo.get('widget_kwargs', {})
                        m_list_key = geo.get('manage_list_key')
                        cont, w = self.create_draggable_container(self.entry_inner_frame, label, cls, key, manage_list_key=m_list_key, **kwargs)
                    elif key.startswith('vehicle_inspection_'):
                        self._loading_memos.append(key)
                        data = geo.get('vehicle_data', {})
                        self.add_vehicle_inspection_box(initial_data=data, key=key)

                # 2. DELAYED RESTORATION: Restore complex states after UI mapped
                def delayed_restore():
                    self.root.update_idletasks()
                    
                    # 1. Restore window state and size
                    try:
                        if config.get('window_state'):
                            if config['window_state'] == 'zoomed':
                                self.root.state('zoomed')
                            else:
                                try:
                                    self.root.state(config['window_state'])
                                except: pass
                        
                        w = config.get('window_width')
                        h = config.get('window_height')
                        if w and h:
                            # Only set geometry if not zoomed
                            if self.root.state() != 'zoomed':
                                self.root.geometry(f"{w}x{h}")
                    except: pass

                    # 2. Restore resolution lock if active
                    if self.resolution_locked:
                        self.root.resizable(False, False)
                        if hasattr(self, 'btn_resolution_lock'):
                            self.btn_resolution_lock.config(text="🔒 해상도 고정됨")

                    # 3. Restore draggable components regardless of lock state
                    # [FIX] Always restore positions if they exist, so movement is persisted 
                    # even if the app was closed while 'unlocked'.
                    draggable_geos = config.get('draggable_geometries', {})
                    for key, geo in draggable_geos.items():
                        if key in self.draggable_items:
                            widget = self.draggable_items[key]
                            try:
                                if geo.get('hidden'):
                                    # [STABILITY] If it's a core widget, ignore hidden status and reset it
                                    if key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
                                        print(f"LOADER: Recovering hidden core widget: {key}")
                                        self.reset_widget_position(None, widget=widget)
                                        continue
                                        
                                    if widget.winfo_manager() == 'grid':
                                        widget.grid_forget()
                                    elif widget.winfo_manager() == 'place':
                                        widget.place_forget()
                                    continue
                                
                                if geo.get('custom_label') and hasattr(widget, '_label_widget'):
                                    widget._label_widget.config(text=geo['custom_label'])
                                
                                # [FIX] Explicitly un-grid before using place manager
                                if widget.winfo_manager() == 'grid':
                                    if not hasattr(widget, '_original_grid_info'):
                                        widget._original_grid_info = widget.grid_info()
                                    widget.grid_forget()
                                
                                self._ensure_placeholder(widget, width=geo.get('width'), height=geo.get('height'))
                                widget.lift()
                                
                                # [FIX] Enforce minimum dimensions to prevent invisible widgets
                                load_x = geo.get('x', 0)
                                load_y = geo.get('y', 0)
                                load_w = max(100, geo.get('width', 200))
                                load_h = max(50, geo.get('height', 100))
                                
                                widget.place(x=load_x, y=load_y, width=load_w, height=load_h)
                            except Exception as e:
                                print(f"Error placing widget {key}: {e}")
                                
                    # 3.5 [SAFEGUARD] Ensure all core widgets are visible even if missing from config
                    for key in getattr(self, 'CORE_DRAGGABLE_KEYS', []):
                        if key in self.draggable_items:
                            widget = self.draggable_items[key]
                            # If not managed by place or grid, or hidden, force reset
                            if widget.winfo_manager() == '' or key not in draggable_geos:
                                print(f"LOADER: Safeguard - Recovering missing/unmanaged core widget: {key}")
                                self.reset_widget_position(None, widget=widget)

                    # 4. Restore sash positions
                    def apply_sashes():
                        try:
                            # Daily usage sash
                            daily_sash = config.get('daily_usage_sash_pos')
                            if daily_sash is not None and hasattr(self, 'daily_usage_paned'):
                                self.daily_usage_paned.sashpos(0, int(daily_sash))
                            
                            # History sash
                            history_sash = config.get('daily_history_sash_pos')
                            if history_sash is not None and hasattr(self, 'daily_history_paned'):
                                self.daily_history_paned.sashpos(0, int(history_sash))
                                
                            # If sash lock is active (check config heavily)
                            config_locked = config.get('daily_usage_sash_locked', False)
                            if config_locked:
                                self.daily_usage_sash_locked = True
                                if hasattr(self, 'btn_sash_lock'):
                                    self.btn_sash_lock.config(text="🔒 경계 고정됨")
                                    self.btn_sash_lock.configure(style="SashLock.TButton")
                                self._start_sash_monitor()
                            elif self.daily_usage_sash_locked: # If self says locked but config didn't (fallback)
                                self._start_sash_monitor()
                        except: pass

                    self.root.after(100, apply_sashes)
                    self.root.after(800, apply_sashes)

                    # 5. [FIX] Recalculate entry_inner_frame height instead of using potentially stale config
                    if hasattr(self, 'entry_inner_frame'):
                        self._adjust_parent_height(self.entry_inner_frame, force=True)
                        # Also refresh canvas
                        self._ensure_canvas_scroll_region()
                    
                    # 6. Final UI refresh
                    for l_key in ['users', 'sites', 'equipments', 'warehouses', 'worktimes']:
                        self.refresh_ui_for_list_change(l_key)

                    # 7. Mark as ready for future saves AFTER all restoration is done
                    def finalize_loading():
                        self.is_ready = True
                        print("APP READY: State restoration complete.")
                    
                    self.root.after(300, finalize_loading) 
                
                # Execute delayed restoration
                self.root.after(300, delayed_restore)

        except Exception as e:
            print(f"Failed to load tab config: {e}")
            # Ensure we eventually become ready even if load failed, to allow new saves
            self.root.after(2000, lambda: setattr(self, 'is_ready', True))
    
    def on_closing(self):
        """Handle window closing event"""
        self.save_tab_config()
        self.root.destroy()
    
    def export_stock_to_excel(self):
        """Export current stock data to Excel"""
        try:
            # Get current filtered data from treeview
            stock_data = []
            
            for item in self.stock_tree.get_children():
                values = self.stock_tree.item(item, 'values')
                stock_data.append({
                    'ID': values[0],
                    '회사코드': values[1],
                    '관리품번': values[2],
                    '품목명': values[3],
                    'SN': values[4],
                    '창고': values[5],
                    '모델명': values[6],
                    '규격': values[7],
                    '품목군코드': values[8],
                    '공급업체': values[9],
                    '제조사': values[10],
                    '제조국': values[11],
                    '가격': values[12],
                    '관리단위': values[13],
                    '수량': values[14],
                    '재고하한': values[15]
                })
            
            if not stock_data:
                messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
                return
            
            # Prepare filename with current date
            current_date = datetime.datetime.now().strftime('%Y%m%d')
            filename = f"재고현황_{current_date}.xlsx"
            
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=filename,
                title="재고 현황 저장",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if save_path:
                stock_df = pd.DataFrame(stock_data)
                stock_df = self.clean_df_export(stock_df)
                self.save_df_to_excel_autofit(stock_df, save_path, "재고현황")
                messagebox.showinfo("완료", f"재고 현황이 저장되었습니다.\n저장 위치: {save_path}")
                
        except Exception as e:
            messagebox.showerror("오류", f"내보내기 실패: {e}")

    def enable_tree_column_drag(self, tree):
        """Ultra-robust drag-and-drop column reordering using root-level event tracking"""
        tree._drag_info = {"col_id": None, "col_name": None, "start_x": 0, "active": False, "ghost": None}
        
        def on_press(event):
            region = tree.identify_region(event.x, event.y)
            if region == "heading":
                col_id = tree.identify_column(event.x)
                col_name = self._get_column_name_from_id(tree, col_id)
                if col_name:
                    tree._drag_info.update({
                        "col_id": col_id,
                        "col_name": col_name,
                        "start_x": event.x_root,
                        "active": False,
                        "motion_id": self.root.bind("<B1-Motion>", on_motion, add="+"),
                        "release_id": self.root.bind("<ButtonRelease-1>", on_release, add="+")
                    })
                
        def on_motion(event):
            di = tree._drag_info
            if not di["col_name"]: return
            
            # Start drag if moved > 5 pixels
            if not di["active"] and abs(event.x_root - di["start_x"]) > 5:
                di["active"] = True
                tree.config(cursor="fleur")
                
                # Visual ghost label
                try:
                    if di["ghost"]: di["ghost"].destroy()
                    di["ghost"] = tk.Toplevel(self.root)
                    di["ghost"].overrideredirect(True)
                    di["ghost"].attributes("-alpha", 0.8)
                    di["ghost"].attributes("-topmost", True)
                    
                    lbl = tk.Label(di["ghost"], text=f" ↔ {di['col_name']} 이동 중 ", 
                                   bg="#424242", fg="white", relief="raised", borderwidth=1, 
                                   padx=8, pady=4, font=('Malgun Gothic', 10, 'bold'))
                    lbl.pack()
                except: pass
            
            if di["active"] and di["ghost"]:
                try:
                    # Follow mouse with offset to stay visible
                    di["ghost"].geometry(f"+{event.x_root + 20}+{event.y_root + 10}")
                except: pass

        def on_release(event):
            di = tree._drag_info
            if di["active"] and di["col_name"]:
                # Convert screen X back to tree-relative X
                tree_x = event.x_root - tree.winfo_rootx()
                target_id = tree.identify_column(tree_x)
                
                if target_id and target_id != di["col_id"]:
                    self._reorder_tree_columns(tree, di["col_id"], target_id)
            
            # Cleanup all bindings and state
            tree.config(cursor="")
            if di["ghost"]:
                try: di["ghost"].destroy()
                except: pass
                di["ghost"] = None
            
            # Unbind specific root-level drag listeners securely
            try:
                if di.get("motion_id"):
                    self.root.unbind("<B1-Motion>", di["motion_id"])
                if di.get("release_id"):
                    self.root.unbind("<ButtonRelease-1>", di["release_id"])
            except: pass
            
            di.update({"col_id": None, "col_name": None, "active": False, "motion_id": None, "release_id": None})

        # Primary entry point: tree-level press
        tree.bind("<Button-1>", on_press, add="+")
        # Direct right-click menu as reliable alternative
        tree.bind("<Button-3>", lambda e: self._show_heading_context_menu(e, tree), add="+")

    def _show_heading_context_menu(self, event, tree):
        """Show context menu on Treeview header to move columns"""
        region = tree.identify_region(event.x, event.y)
        if region == "heading":
            column_id = tree.identify_column(event.x)
            
            # Identify the column name from visual ID
            col_name = self._get_column_name_from_id(tree, column_id)
            if not col_name: return

            menu = tk.Menu(self.root, tearoff=0)
            menu.add_command(label=f"⬅️ '{col_name}' 왼쪽으로 이동", 
                             command=lambda: self._move_column_visual(tree, column_id, -1))
            menu.add_command(label=f"➡️ '{col_name}' 오른쪽으로 이동", 
                             command=lambda: self._move_column_visual(tree, column_id, 1))
            menu.add_separator()
            menu.add_command(label="⚙️ 컬럼 관리(숨기기/보이기)...", command=self.show_column_visibility_dialog)
            
            menu.post(event.x_root, event.y_root)

    def _get_column_name_from_id(self, tree, column_id):
        """Helper to get actual column name from visual ID like #1"""
        try:
            if not column_id or not column_id.startswith("#"): return None
            # Extract N from #N
            vis_idx = int(column_id[1:]) - 1
            
            # Get displaycolumns accurately
            disp_tuple = tree.cget("displaycolumns")
            if not disp_tuple or disp_tuple == ("#all",) or disp_tuple == ["#all"] or disp_tuple == "":
                visible_names = list(tree.cget("columns"))
            else:
                visible_names = list(disp_tuple)
            
            if 0 <= vis_idx < len(visible_names):
                return visible_names[vis_idx]
        except: pass
        return None

    def _move_column_visual(self, tree, source_id, direction):
        """Move a column left (-1) or right (1) based on its current position"""
        col_name = self._get_column_name_from_id(tree, source_id)
        if not col_name: return
        
        try:
            disp_tuple = tree.cget("displaycolumns")
            if not disp_tuple or disp_tuple == ("#all",) or disp_tuple == ["#all"] or disp_tuple == "":
                visible_names = list(tree.cget("columns"))
            else:
                visible_names = list(disp_tuple)

            if col_name not in visible_names: return
            
            src_idx = visible_names.index(col_name)
            tgt_idx = src_idx + direction
            
            if 0 <= tgt_idx < len(visible_names):
                visible_names.pop(src_idx)
                visible_names.insert(tgt_idx, col_name)
                tree["displaycolumns"] = visible_names
                
                # Update persistence
                if tree == self.daily_usage_tree:
                    self.manual_visible_cols = visible_names
                    self.save_tab_config()
                elif tree == self.monthly_usage_tree:
                    self.monthly_visible_cols = visible_names
                    self.save_tab_config()
        except: pass

    def _reorder_tree_columns(self, tree, source_id, target_id):
        """Internal helper to swap displaycolumns for drag-and-drop"""
        src_name = self._get_column_name_from_id(tree, source_id)
        tgt_name = self._get_column_name_from_id(tree, target_id)
        
        if not src_name or not tgt_name or src_name == tgt_name: return
        
        try:
            disp_tuple = tree.cget("displaycolumns")
            if not disp_tuple or disp_tuple == ("#all",) or disp_tuple == ["#all"] or disp_tuple == "":
                visible_names = list(tree.cget("columns"))
            else:
                visible_names = list(disp_tuple)
            
            if src_name in visible_names and tgt_name in visible_names:
                src_idx = visible_names.index(src_name)
                tgt_idx = visible_names.index(tgt_name)
                
                # Strategic insertion: move before or after target 
                # (Simple swap approach: just pop and insert)
                visible_names.pop(src_idx)
                visible_names.insert(tgt_idx, src_name)
                
                tree["displaycolumns"] = visible_names
                
                # Update persistence
                if tree == self.daily_usage_tree:
                    self.manual_visible_cols = visible_names
                    self.save_tab_config()
                elif tree == self.monthly_usage_tree:
                    self.monthly_visible_cols = visible_names
                    self.save_tab_config()
        except: pass


    def treeview_sort_column(self, tv, col, reverse):
        """Sort treeview contents when a column header is clicked"""
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        
        # Separate the 'Total' row (tagged 'total') from sorting
        data_rows = []
        total_row = None
        
        for val, k in l:
            # Check tags for this item
            tags = tv.item(k, 'tags')
            # If tags is a tuple/list, check if 'total' is in it. 
            # If it's a string (though usually tuple), check equality or contains.
            if tags and ('total' in tags or tags == 'total'):
                total_row = (val, k)
            else:
                data_rows.append((val, k))
                
        # Helper for numeric conversion
        def convert(val):
            try:
                # Remove common non-numeric chars
                s = str(val).replace(',', '').replace('시간', '').replace('Hrs', '').replace('원', '').replace(' ', '').replace('(', '').replace(')', '')
                if not s: return 0.0
                return float(s)
            except ValueError:
                return str(val).lower() # Default to string sort

        try:
            data_rows.sort(key=lambda t: convert(t[0]), reverse=reverse)
        except Exception:
            data_rows.sort(key=lambda t: t[0].lower(), reverse=reverse)

        # Rearrange items in sorted positions
        for index, (val, k) in enumerate(data_rows):
            tv.move(k, '', index)
            
        # Ensure Total row is always at the bottom
        if total_row:
             tv.move(total_row[1], '', 'end')

        # Reverse sort next time
        tv.heading(col, command=lambda: self.treeview_sort_column(tv, col, not reverse))

    def switch_to_daily_usage_details(self):
        """Switch to Daily Usage tab and apply current filters from Sales tab"""
        # [FIX] Aggressively hide any active suggestion window before switching
        if hasattr(self, '_suggestion_windows'):
            for win in self._suggestion_windows.values():
                win.hide()
        if hasattr(self, 'cb_sales_filter_site') and hasattr(self.cb_sales_filter_site, '_suggestion_win'):
            self.cb_sales_filter_site._suggestion_win.hide()

        site = self.cb_sales_filter_site.get().strip()
        start = self.sales_start_date.get_date()
        end = self.sales_end_date.get_date()
        
        # 1. Switch Tab (index 5)
        self.notebook.select(5)
        
        # [NEW] Force focus transition
        if hasattr(self, 'daily_usage_tree'):
            self.daily_usage_tree.focus_set()

        # 2. Sync Filters
        if hasattr(self, 'cb_daily_filter_site'):
            self.cb_daily_filter_site.set(site)
        
        if hasattr(self, 'ent_daily_start_date'):
            self.ent_daily_start_date.set_date(start)
            
        if hasattr(self, 'ent_daily_end_date'):
            self.ent_daily_end_date.set_date(end)
            
        # 3. Trigger Search
        self.update_daily_usage_view()


if __name__ == "__main__":
    root = tk.Tk()
    app = MaterialManager(root)
    root.mainloop()
