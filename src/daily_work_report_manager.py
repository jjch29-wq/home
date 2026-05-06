import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.cell.cell import MergedCell
import os
import datetime

class DailyWorkReportManager:
    def __init__(self, template_path):
        self.template_path = template_path
        # Default mapping for fallback
        self.default_mapping = {
            'header': {'date': 'F2'},
            'general': {
                'company': 'E5', 'project_name': 'E6', 'standard': 'E7', 'equipment': 'E8',
                'report_no': 'K5', 'inspection_item': 'K6', 'inspector': 'K7', 'inspector_n8': 'N8', 'car_no': 'N9'
            },
            'methods': {
                'RT': {'row': 13}, 'UT': {'row': 14}, 'MT': {'row': 15}, 'PT': {'row': 16},
                'HT': {'row': 17}, 'VT': {'row': 18}, 'LT': {'row': 19}, 'ET': {'row': 20},
                'PAUT': {'row': 21}
            },
            'rtk': {
                'center_miss': 'D34', 'density': 'F34', 'marking_miss': 'H34', 'film_mark': 'J34',
                'handling': 'L34', 'customer_complaint': 'N34', 'etc': 'P34', 'total': 'R34'
            },
            'ot': {
                'row1_name': 'B38', 'row1_company': 'F38', 'row1_method': 'I38', 'row1_hours': 'K38', 'row1_amount': 'N38',
                'row2_name': 'B39', 'row2_company': 'F39', 'row2_method': 'I39', 'row2_hours': 'K39', 'row2_amount': 'N39'
            },
            'materials': {
                'RT T200': 43, 'RT AA400': 44, 'RT Other': 45,
                'MT WHITE': 46, 'MT 7C-BLACK': 47,
                'PT Penetrant': 48, 'PT Cleaner': 49, 'PT Developer': 50
            },
            'vehicles': {
                'row_start': 26, 
                'col_map': {'vehicle_info': 'B', 'mileage': 'H', 'fuel': 'C', 'clean': 'E', 'oil': 'G', 'tire': 'I', 'light': 'K', 'safety': 'M'}
            }
        }

    def generate_report(self, data, output_path, custom_mapping=None):
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import re as _re
        from copy import copy

        custom_mapping = custom_mapping or {}
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template not found at {self.template_path}")

        wb = openpyxl.load_workbook(self.template_path)
        sheet = wb.active
        mapping = custom_mapping if custom_mapping else self.default_mapping

        def safe_write(cell_coord, value, is_currency=False, is_bold=False):
            if not cell_coord: return
            try:
                cell = sheet[cell_coord]
                if isinstance(cell, MergedCell):
                    for m_range in sheet.merged_cells.ranges:
                        if cell_coord in m_range:
                            cell = sheet.cell(row=m_range.min_row, column=m_range.min_col)
                            break
                cell.value = value
                cell.font = Font(name='맑은 고딕', size=9, bold=is_bold)
                # [FIX] shrinkToFit and wrapText are mutually exclusive in Excel - use wrapText only
                cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                if is_currency:
                    cell.number_format = '#,##0 "원"'
            except: pass

        # 1. Header & General
        d = data.get('date', datetime.date.today())
        weekdays = ["월", "화", "수", "목", "금", "토", "일"]
        date_str = f"{d.year}년 {d.month}월 {d.day}일 ({weekdays[d.weekday()]})"
        safe_write(mapping.get('header', {}).get('date', 'F2'), date_str)

        gen_map = mapping.get('general', {})
        for key in ['company', 'project_name', 'standard', 'equipment', 'report_no', 'inspection_item', 'inspector', 'car_no', 'inspector_n8']:
            val_key = 'inspector' if key == 'inspector_n8' else key
            safe_write(gen_map.get(key), data.get(val_key, ''))

        # 2. Methods (Dynamic Offset)
        methods_data = data.get('methods', {})
        current_methods = list(methods_data.keys())
        method_offset = max(0, len(current_methods) - 4)
        
        import re as _re
        def fix_c(coord, offset):
            if not coord or not isinstance(coord, str): return coord
            match = _re.match(r"([A-Z]+)([0-9]+)", coord)
            if match:
                col, row = match.groups()
                row_val = int(row)
                if row_val >= 18: return f"{col}{row_val + offset}"
            return coord

            # [UNIFIED HEIGHT BALANCE] No row deletion for methods either
            # We will compensate for added method rows by reducing Note heights later
            pass

        method_base_rows = [13, 14, 15, 16]
        if method_offset > 0:
            sheet.insert_rows(17, method_offset)
            for i in range(method_offset):
                target_r = 18 + method_offset + i
                if target_r <= 25 + method_offset:
                    sheet.row_dimensions[target_r].height = 5
            method_base_rows += list(range(17, 17 + method_offset))

        for idx, m_name in enumerate(current_methods):
            if idx < len(method_base_rows):
                r = method_base_rows[idx]; m = methods_data[m_name]
                safe_write(f'B{r}', m_name); safe_write(f'E{r}', m.get('unit', ''))
                safe_write(f'H{r}', m.get('qty', 0)); safe_write(f'K{r}', m.get('price', 0), is_currency=True)
                safe_write(f'N{r}', m.get('travel', 0), is_currency=True); safe_write(f'Q{r}', m.get('total', 0), is_currency=True)

        # 3. Section 2 (Notes) - Styling & Normalize
        # [NEW] Sync Row 17 height with Row 10 height for consistency
        h10 = sheet.row_dimensions[10].height if sheet.row_dimensions[10].height else 15
        sheet.row_dimensions[17].height = h10
        for m_range in list(sheet.merged_cells.ranges):
            if 18 + method_offset <= m_range.min_row and m_range.max_row <= 25 + method_offset:
                try: sheet.unmerge_cells(m_range.coord)
                except: pass
        try: sheet.merge_cells(f"B{18+method_offset}:S{25+method_offset}")
        except: pass
        
        # 4. Section 4 (RTK)
        rtk_data = data.get('rtk_quality', {}); rtk_map = mapping.get('rtk', {})
        rtk_key_map = {'center_miss':'센터미스','density':'농도','marking_miss':'마킹미스','film_mark':'필름마크','handling':'취급부주의','customer_complaint':'고객불만','etc':'기타'}
        for eng, kor in rtk_key_map.items():
            safe_write(fix_c(rtk_map.get(eng), method_offset), rtk_data.get(kor, 0))
        total_rtk = rtk_data.get('총계', sum([v for v in rtk_data.values() if isinstance(v, (int, float))]))
        safe_write(fix_c(rtk_map.get('total'), method_offset), total_rtk)

        # 5. Section 5 (OT)
        for r_search in range(1, 100):
            if sheet.cell(row=r_search, column=2).value == "검사자":
                ot_header_row = r_search; break
        else: ot_header_row = 38 + method_offset
        ot_list = data.get('ot_status', [])
        ot_base_count = 2  # 템플릿 기본 OT 행 수
        ot_extra = max(0, len(ot_list) - ot_base_count)  # [FIX] 초과 인원 수 계산

        # [FIX] OT 인원이 2명 초과 시 행 삽입
        if ot_extra > 0:
            ot_insert_row = ot_header_row + ot_base_count + 1  # 기존 2개 행 바로 아래
            print(f"DEBUG: Inserting {ot_extra} OT rows at row {ot_insert_row}")
            sheet.insert_rows(ot_insert_row, ot_extra)
            from copy import copy
            for r_new in range(ot_insert_row, ot_insert_row + ot_extra):
                sheet.row_dimensions[r_new].height = sheet.row_dimensions[ot_header_row + 1].height or 15
                for col in range(1, 20):
                    try:
                        source_cell = sheet.cell(row=ot_header_row + 1, column=col)
                        target_cell = sheet.cell(row=r_new, column=col)
                        if source_cell.has_style:
                            target_cell.font = copy(source_cell.font)
                            target_cell.border = copy(source_cell.border)
                            target_cell.fill = copy(source_cell.fill)
                            target_cell.alignment = copy(source_cell.alignment)
                            target_cell.number_format = source_cell.number_format
                    except: pass

        for i, ot in enumerate(ot_list):
            r = ot_header_row + 1 + i
            safe_write(f"B{r}", ot.get('names', '')); safe_write(f"F{r}", ot.get('company', ''))
            safe_write(f"I{r}", ot.get('method', '')); safe_write(f"K{r}", ot.get('ot_hours', ''))
            safe_write(f"N{r}", ot.get('ot_amount', ''), is_currency=True)

        # 6. Section 6 (Materials)
        materials_data = data.get('materials', {}); active_rt = []
        for k, v in materials_data.items():
            if isinstance(v, dict) and (k.startswith('RT ') or k.startswith('RT_ROW_') or v.get('is_rt')):
                name = v.get('name', '').strip()
                if not name: continue
                item = v.copy(); item['name'] = name; active_rt.append(item)
        
        # RT dynamic rows: if more than 4, insert additional rows
        rt_count = len(active_rt)
        base_rt_limit = 4
        rt_extra = max(0, rt_count - base_rt_limit)
        
        if rt_extra > 0:
            print(f"DEBUG: Inserting {rt_extra} rows for RT extra materials at row {46 + ot_extra}")
            # [FIX] OT 행 삽입 이후이므로 RT 삽입 위치도 ot_extra만큼 이동
            sheet.insert_rows(46 + ot_extra, rt_extra)
            from copy import copy
            for r_new in range(46 + ot_extra, 46 + ot_extra + rt_extra):
                print(f"DEBUG: Copying style from Row {45 + ot_extra} to New Row {r_new}")
                sheet.row_dimensions[r_new].height = 30
                for col in range(1, 20): # Columns A to S
                    try:
                        source_cell = sheet.cell(row=45 + ot_extra, column=col)
                        target_cell = sheet.cell(row=r_new, column=col)
                        if source_cell.has_style:
                            target_cell.font = copy(source_cell.font)
                            target_cell.border = copy(source_cell.border)
                            target_cell.fill = copy(source_cell.fill)
                            target_cell.alignment = copy(source_cell.alignment)
                            target_cell.number_format = source_cell.number_format
                    except: pass
            # We do NOT delete rows from Notes, we will adjust their height later
        
        total_offset = method_offset + ot_extra + rt_extra  # [FIX] OT 초과 행 포함
        
        # 1. Category Merges (B:C) Helper
        def safe_merge(sheet, s_row, s_col, e_row, e_col, value):
            for m_range in list(sheet.merged_cells.ranges):
                if not (e_row < m_range.min_row or s_row > m_range.max_row or 
                        e_col < m_range.min_col or s_col > m_range.max_col):
                    try: sheet.unmerge_cells(m_range.coord)
                    except: pass
            try:
                sheet.merge_cells(start_row=s_row, start_column=s_col, end_row=e_row, end_column=e_col)
                cell = sheet.cell(row=s_row, column=s_col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
            except: pass

        rt_display_count = max(1, rt_count)
        rt_end_row = 42 + method_offset + ot_extra + max(1, rt_count)  # [FIX] method+OT+RT 모두 반영
        
        safe_merge(sheet, 43 + method_offset + ot_extra, 2, rt_end_row, 3, "RT") 
        safe_merge(sheet, 47 + total_offset, 2, 48 + total_offset, 3, "MT")
        safe_merge(sheet, 49 + total_offset, 2, 51 + total_offset, 3, "PT")

        # Set column widths
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 12

        # 2. Item Name (D:E) and Spec (F:G) Horizontal Merges
        # We process all rows from 43 to 51 + rt_extra
        for r in range(43, 52 + rt_extra):
            try:
                # Clean up existing merges first
                for m_range in list(sheet.merged_cells.ranges):
                    if m_range.min_row == r and m_range.max_row == r and (m_range.min_col >= 4 and m_range.max_col <= 7):
                        sheet.unmerge_cells(m_range.coord)
                
                if r <= 42 + rt_display_count:
                    # RT Area: Separate Name (D:E) and Spec (F:G)
                    sheet.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
                    sheet.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
                else:
                    # MT/PT Area: Wide Combined Name/Spec (D:G)
                    sheet.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
                    
                # [HORIZONTAL MERGES] Apply to ALL material rows (RT, MT, PT)
                # Merges for: I:J, K:L, M:N, O:P, Q:S
                for s_c, e_c in [(9,10), (11,12), (13,14), (15,16), (17,19)]:
                    try:
                        # Clean existing merges in this row/column range
                        for m_range in list(sheet.merged_cells.ranges):
                            if m_range.min_row == r and m_range.max_row == r and m_range.min_col == s_c:
                                sheet.unmerge_cells(m_range.coord)
                        sheet.merge_cells(start_row=r, start_column=s_c, end_row=r, end_column=e_c)
                    except: pass
            except: pass

        # [DYNAMIC STYLE COPY & DATA PROTECTION]
        # 1. Clear the RT Range first to remove template remnants
        # [FIX] mat_start reflects method + OT offsets
        mat_start = 43 + method_offset + ot_extra
        for r in range(mat_start, mat_start + 4 + rt_extra):
            for col in range(4, 20): # D to S
                cell = sheet.cell(row=r, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = None

        # Helper: Remove known brand prefixes for display
        def strip_brand(name):
            for prefix in ['Carestream ', 'AGFA ', 'Fuji ', 'Kodak ']:
                if name.startswith(prefix):
                    return name[len(prefix):]
            return name

        # Write RT Data with Standardized Formatting
        print(f"DEBUG: Starting RT Data Writing Loop. Total RT items: {len(active_rt)}, mat_start={mat_start}")
        for idx, m in enumerate(active_rt):
            r = mat_start + idx  # [FIX] mat_start 사용
            if r > mat_start + 3 + rt_extra:  # [FIX] 범위 체크도 mat_start 기준
                print(f"DEBUG: Breaking RT loop at r={r}")
                break
            
            # Strip brand prefix so short name fits in merged cell
            display_name = strip_brand(m.get('name', ''))
            print(f"DEBUG: Writing RT Item {idx+1} to Row {r}: {display_name}")
            safe_write(f'D{r}', display_name)
            # Apply style AFTER safe_write
            cell_name = sheet.cell(row=r, column=4)
            cell_name.font = Font(name='맑은 고딕', size=9)
            cell_name.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            sheet.row_dimensions[r].height = 30
            
            safe_write(f'F{r}', m.get('spec', ''))
            sheet.cell(row=r, column=6).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            safe_write(f'H{r}', m.get('unit', '매'))
            
            # Correct Mapping: K=Out, M=Used, O=In (Q is intentionally left blank)
            safe_write(f'K{r}', '-')
            used_val = int(m.get('used', 0))
            in_val = int(m.get('in', 0))
            safe_write(f'M{r}', f"{used_val} 매") # Usage is M
            safe_write(f'O{r}', f"{in_val} 매")   # Incoming is O
            # Stock (Q) removed per request

        # [SELECTIVE CONTENT CLEARING & SAFETY-MARGIN COMPENSATION]
        # Clear template text in B18:S25 to allow clean compression
        for r_clear in range(18 + method_offset, 26 + method_offset):
            for c_clear in range(2, 20): # B to S
                cell = sheet.cell(row=r_clear, column=c_clear)
                if not isinstance(cell, MergedCell):
                    cell.value = None

        # [HARD RESET FORMATTING & HIERARCHY]
        tiny_font = Font(size=1)
        no_wrap = Alignment(wrap_text=False, horizontal='center', vertical='center')
        
        rt_count = len(active_rt)
        rt_extra = max(0, rt_count - 4)
        
        # [FINAL STABLE COMPRESSION & OFFSET CALCULATION]
        # [FIX] ot_extra must be included
        total_offset = method_offset + ot_extra + rt_extra  # CRITICAL for Section 6 mapping
        
        # [STABLE COMPRESSION: CLEAR & RESIZE (SAFE)]
        # We avoid delete_rows as it destroys template merges in Sections 3-6.
        note_range_start = 18 + method_offset
        note_range_end = 25 + method_offset
        
        # 1. Clear contents and unmerge ONLY in the note range
        for r in range(note_range_start, note_range_end + 1):
            # Reset Row Height
            sheet.row_dimensions[r].height = 6.0
            sheet.row_dimensions[r].custom_height = True
            
            # Clear Cells and Alignment
            for c in range(1, 20): # A to S
                cell = sheet.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.value = None
                    cell.alignment = Alignment(wrap_text=False, vertical='center', horizontal='center')
                    cell.font = Font(size=1)
                
        # 2. Kill merges in this specific range to allow shrinking
        merges_to_kill = []
        for m_range in list(sheet.merged_cells.ranges):
            if m_range.min_row >= note_range_start and m_range.max_row <= note_range_end:
                merges_to_kill.append(m_range)
        for m in merges_to_kill:
            try: sheet.unmerge_cells(str(m))
            except: pass
            
        # 3. Method Rows (13-17 + inserted)
        for r in range(13, 18 + method_offset):
            sheet.row_dimensions[r].height = 15.0
            sheet.row_dimensions[r].custom_height = True
            
        # 4. Materials (Section 6) prominent height
        mat_start_fixed = 43 + method_offset + ot_extra  # [FIX] RT 시작 행에 OT offset 반영
        for r in range(mat_start_fixed, mat_start_fixed + rt_count):
            sheet.row_dimensions[r].height = 25.0
            sheet.row_dimensions[r].custom_height = True
            
        # 5. Shrink Spacer Row 42
        spacer_row = 42 + method_offset + ot_extra  # [FIX] spacer row도 OT offset 반영
        sheet.row_dimensions[spacer_row].height = 5
        sheet.row_dimensions[spacer_row].custom_height = True

        # 3. Hide unused RT rows (Only if count < 4)
        if rt_count < 4:
            for r in range(mat_start + rt_count, mat_start + 4):  # [FIX] mat_start 기준으로 계산
                sheet.row_dimensions[r].height = 0
                sheet.row_dimensions[r].hidden = True
                # Clear content/borders
                for col in range(1, 20):
                    cell = sheet.cell(row=r, column=col)
                    if not isinstance(cell, MergedCell):
                        cell.value = None
                        cell.border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

        # Add placeholders ONLY to active RT rows that are empty
        for r in range(mat_start, mat_start + rt_count):  # [FIX] mat_start 기준
            if r >= mat_start + len(active_rt):
                safe_write(f'K{r}', '-'); safe_write(f'M{r}', '매'); safe_write(f'O{r}', '0 매')
            
        synonyms = {
            'MT WHITE': ['MT WHITE', '백색', '백색자분', 'WHITE'],
            'MT 7C-BLACK': ['MT 7C-BLACK', '7C', '자분', 'BLACK', '7C-BLACK'],
            'PT Penetrant': ['PT Penetrant', '침투', '침투액', 'Penetrant'],
            'PT Cleaner': ['PT Cleaner', '세척', '세척액', 'Cleaner'],
            'PT Developer': ['PT Developer', '현상', '현상액', '현상제', 'Developer']
        }
        # [DYNAMIC OFFSET] MT: 47+, PT: 49+
        mat_map = {
            'MT WHITE': 47 + total_offset, 
            'MT 7C-BLACK': 48 + total_offset, 
            'PT Penetrant': 49 + total_offset, 
            'PT Cleaner': 50 + total_offset, 
            'PT Developer': 51 + total_offset
        }
        
        # Korean Display Names for MT/PT
        display_names = {
            'MT WHITE': '백색페인트', 'MT 7C-BLACK': '흑색자분', 
            'PT Penetrant': '침투액', 'PT Cleaner': '세척액', 'PT Developer': '현상액'
        }
        
        # Category mapping for Column B
        categories = {
            'MT WHITE': 'MT', 'MT 7C-BLACK': 'MT',
            'PT Penetrant': 'PT', 'PT Cleaner': 'PT', 'PT Developer': 'PT'
        }
        
        print(f"DEBUG: MT/PT Write. total_offset={total_offset}, rt_extra={rt_extra}, method_offset={method_offset}")
        print(f"DEBUG: mat_map={mat_map}")
        
        for m_key, r in mat_map.items():
            # B column ("MT"/"PT") is already handled by safe_merge above.
            # Only write specific name to D:G area here.
            display_name = display_names.get(m_key, m_key)
            print(f"DEBUG: Writing '{display_name}' to D{r}")
            safe_write(f'D{r}', display_name) # Specific name in D:G

            # Merge D:G for Specific Item Name
            try:
                for m_range in list(sheet.merged_cells.ranges):
                    if m_range.min_row == r and m_range.max_row == r and (m_range.min_col >= 4 and m_range.max_col <= 7):
                        sheet.unmerge_cells(m_range.coord)
                sheet.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
            except Exception as e:
                print(f"DEBUG: Merge error at row {r}: {e}")

            # ALWAYS write the unit to H (All MT/PT use CAN)
            safe_write(f'H{r}', "CAN")
            
            m_d = {}
            for syn in synonyms.get(m_key, []):
                m_d = data.get('materials', {}).get(syn)
                if m_d: break
            
            if m_d:
                # K=Out, M=Used, O=In (Q is intentionally left blank)
                safe_write(f'K{r}', '-'); safe_write(f'M{r}', int(m_d.get('used', 0)))
                safe_write(f'O{r}', int(m_d.get('in', 0)))

        # [BORDER & ALIGNMENT] Apply THIN border to B42:S(51 + total_offset)
        thin = Side(style='thin')
        for r in range(42, 52 + total_offset):
            for col in range(2, 20): # B(2) to S(19)
                cell = sheet.cell(row=r, column=col)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # [MT/PT WRITE - AFTER BORDER LOOP to prevent overwrite]
        # Write AFTER borders are applied so values are final
        for m_key, r in mat_map.items():
            display_name = display_names.get(m_key, m_key)
            # Write directly to the cell, bypassing safe_write to avoid wrapText interference
            cell = sheet.cell(row=r, column=4)  # Column D
            cell.value = display_name
            cell.font = Font(name='맑은 고딕', size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            print(f"DEBUG: FINAL WRITE '{display_name}' to D{r}")
            
            safe_write(f'H{r}', "CAN")
            m_d = {}
            for syn in synonyms.get(m_key, []):
                m_d = data.get('materials', {}).get(syn)
                if m_d: break
            if m_d:
                safe_write(f'K{r}', '-')
                safe_write(f'M{r}', int(m_d.get('used', 0)))
                safe_write(f'O{r}', int(m_d.get('in', 0)))

        # [EXCEL PRINT SETUP] Standard A4 Setup & Page Break Reset
        # (Fit to Page removed to allow manual height shrinking to be visible)
        print(f"DEBUG: Resetting Page Breaks and Setting Print Area to A1:S{51 + total_offset}")
        
        # 1. Clear all existing manual row breaks safely
        # sheet.row_breaks is a RowBreak object, we must clear its 'brk' list
        sheet.row_breaks.brk = [] 
        # 2. Update Print Area to include all shifted rows
        sheet.print_area = f'A1:S{51 + total_offset}'
        
        # 3. Add a new page break ONLY at the very end of our dynamic content
        from openpyxl.worksheet.pagebreak import Break
        sheet.row_breaks.append(Break(id=51 + total_offset))
        
        # 4. Standard A4 Settings
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToHeight = None
        sheet.page_setup.fitToWidth = None
        sheet.sheet_properties.pageSetUpPr.fitToPage = False

        # Final Cleanup: Section 3 (Vehicles) & B27 Cleanup
        veh_list = data.get('vehicles', []); v = veh_list[0] if veh_list else {}
        veh_row = 27 + method_offset
        safe_write(f"B{veh_row}", ""); safe_write(f"H{veh_row}", v.get('mileage', ''))
        
        chk_rows = {'out': 29 + total_offset, 'in': 30 + total_offset}
        for rs in range(25+total_offset, 35+total_offset):
            val = sheet.cell(row=rs, column=2).value
            if val == "출차시": chk_rows['out'] = rs
            if val == "입차시": chk_rows['in'] = rs
        
        chk_map = {'exterior':'E','cleanliness':'H','cleaning':'K','locking':'N'}
        WHITE_SQ, BLACK_SQ = "\u25a1", "\u25a0"
        import re
        for rk, row_idx in chk_rows.items():
            for ck, col_let in chk_map.items():
                val = v.get(f"{rk}_{ck}")
                if val:
                    coord = f"{col_let}{row_idx}"; cell = sheet[coord]
                    if isinstance(cell, MergedCell):
                        for mr in sheet.merged_cells.ranges:
                            if coord in mr: cell = sheet.cell(row=mr.min_row, column=mr.min_col); break
                    if cell.value and isinstance(cell.value, str):
                        pattern = f"({WHITE_SQ})(\\s*){re.escape(val)}"
                        if re.search(pattern, cell.value):
                            cell.value = re.sub(pattern, f"{BLACK_SQ}\\2{val}", cell.value)

        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.horizontalCentered = True
        sheet.page_setup.verticalCentered = False # Top-aligned for precision
        
        # Exact Margins (Total 0.8" vertical consumption = 57.6pt)
        sheet.page_margins.top = 0.4; sheet.page_margins.bottom = 0.4
        sheet.page_margins.left = 0.3; sheet.page_margins.right = 0.3
        
        # A4 Usable Height at 100% scale is approx 780pt.
        # Let's allocate heights precisely:
        # 1. Fixed Header (1-11): 11 rows * 22pt = 242pt
        for r in range(1, 12): sheet.row_dimensions[r].height = 22
        
        # 2. Title Rows (12, 26, 31, 37, 42): 5 rows * 26pt = 130pt
        # 2. Title Rows (12, 26, 31, 37, 42)
        fixed_titles = [12, 26, 31, 37, 42]
        
        # [ROW 51 FIT OPTIMIZATION]
        # 1. Global Row Height Baseline (Aggressive Slimming)
        for r in range(1, 101):
            try: sheet.row_dimensions[r].height = 14 # Slimmer baseline
            except: pass
        
        # 2. Precision Column Widths (Remain slim)
        slim_widths = {
            'A': 1, 'B': 7, 'C': 1, 'D': 9, 'E': 1, 'F': 9, 'G': 1, 'H': 5, 
            'I': 4, 'J': 4, 'K': 5, 'L': 4, 'M': 5, 'N': 4, 'O': 5, 
            'P': 4, 'Q': 5, 'R': 4, 'S': 10
        }
        for col, width in slim_widths.items():
            sheet.column_dimensions[col].width = width

        # 3. Precision Row Heights (Recalibrated for Row 51 fit)
        # Header (1-11): 11 * 15 = 165pt
        for r in range(1, 12): sheet.row_dimensions[r].height = 15
        # Titles: 5 * 20 = 100pt
        for t_row in [12, 26, 31, 37, 42]:
            try: sheet.row_dimensions[t_row + method_offset].height = 20
            except: pass
        # Section 2 (Notes): SKIP - already set by compression logic above
        # DO NOT overwrite the heights set earlier!
        # for b_row in range(18+method_offset, 26+method_offset):
        #     sheet.row_dimensions[b_row].height = 25
        # Materials: ~9 * 16 = 144pt
        for m_row in range(43 + method_offset, 52 + total_offset):
            sheet.row_dimensions[m_row].height = 16
            
        # 8. Final Print Setup - FIXED 95% SCALE & TIGHT MARGINS
        sheet.sheet_properties.pageSetUpPr.fitToPage = False
        sheet.page_setup.scale = 95
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.horizontalCentered = True
        sheet.page_setup.verticalCentered = True
        
        # Symmetrical Margins with MINIMIZED BOTTOM for extra space
        sheet.page_margins.left = 0.5; sheet.page_margins.right = 0.5
        sheet.page_margins.top = 0.4; sheet.page_margins.bottom = 0.2

        sheet.print_area = f"A1:S{51 + method_offset + rt_extra}"
        
        # [ABSOLUTE LAST STEP: CELL-TARGETED UNMERGE + WRITE]
        from openpyxl.cell.cell import Cell as RealCell
        black_font = Font(name='맑은 고딕', size=10, bold=False, color='000000')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
        
        # Helper 1: free a cell by unmerging (used BEFORE merge_cells)
        def free_cell(sheet, row, col):
            for m_range in list(sheet.merged_cells.ranges):
                if row >= m_range.min_row and row <= m_range.max_row and col >= m_range.min_col and col <= m_range.max_col:
                    sheet.unmerge_cells(m_range.coord)
            key = (row, col)
            if key not in sheet._cells or isinstance(sheet._cells[key], MergedCell):
                sheet._cells[key] = RealCell(sheet, row=row, column=col)
        
        # Helper 2: write to cell via _cells dict WITHOUT unmerging (used AFTER merge_cells)
        thin = Side(style='thin')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        def write_cell(sheet, row, col, value, font=None, align=None):
            key = (row, col)
            if key not in sheet._cells or isinstance(sheet._cells[key], MergedCell):
                sheet._cells[key] = RealCell(sheet, row=row, column=col)
            cell = sheet._cells[key]
            cell.value = value
            cell.border = thin_border
            if font: cell.font = font
            if align: cell.alignment = align
        
        # 1. Category merges (RT, MT, PT)
        rt_end = mat_start + max(1, rt_count) - 1  # [FIX] mat_start 기준으로 RT 끝 행 계산
        merge_plan = [
            (mat_start, 2, rt_end, 3, "RT"),          # [FIX] mat_start 사용
            (47 + total_offset, 2, 48 + total_offset, 3, "MT"),   # [FIX] total_offset 사용
            (49 + total_offset, 2, 51 + total_offset, 3, "PT"),   # [FIX] total_offset 사용
        ]
        for s_row, s_col, e_row, e_col, label in merge_plan:
            # Phase 1: Free all cells (unmerge existing)
            for r in range(s_row, e_row + 1):
                for c in range(s_col, e_col + 1):
                    free_cell(sheet, r, c)
            # Phase 2: Create merge
            try:
                sheet.merge_cells(start_row=s_row, start_column=s_col, end_row=e_row, end_column=e_col)
            except Exception as e:
                print(f"DEBUG: Merge B{s_row}:C{e_row} error: {e}")
            # Phase 3: Write to top-left WITHOUT unmerging
            write_cell(sheet, s_row, s_col, label, black_font, center_align)
            # Phase 4: Apply border to ALL cells in merge range (including MergedCells)
            for r in range(s_row, e_row + 1):
                for c in range(s_col, e_col + 1):
                    key = (r, c)
                    if key not in sheet._cells or isinstance(sheet._cells[key], MergedCell):
                        sheet._cells[key] = RealCell(sheet, row=r, column=c)
                    sheet._cells[key].border = thin_border
            print(f"DEBUG: B{s_row}:C{e_row} = '{label}'")
        
        # 2. Force D column item names (write without touching merges)
        for m_key, r in mat_map.items():
            display_name = display_names.get(m_key, m_key)
            write_cell(sheet, r, 4, display_name, black_font, center_align)
        
        wb.save(output_path); return output_path
