import openpyxl
import os

path = r"c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\Na-aba\home\resources\Template_DailyWorkReport.xlsx"

print(f"Checking: {path}")
if not os.path.exists(path):
    print("File not found.")
else:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active
    # Check Column A (1), B (2), C (3)
    for r in range(1, 100):
        val_a = sheet.cell(row=r, column=1).value
        val_b = sheet.cell(row=r, column=2).value
        if (val_a and any(str(val_a).startswith(s) for s in ["3.", "4.", "5.", "6."])) or \
           (val_b and any(str(val_b).startswith(s) for s in ["3.", "4.", "5.", "6."])):
            print(f"Row {r}: A='{val_a}', B='{val_b}'")
