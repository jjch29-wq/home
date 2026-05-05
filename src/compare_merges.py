import os
import openpyxl
from daily_work_report_manager import DailyWorkReportManager
import datetime

# Setup paths
current_dir = os.path.dirname(os.path.abspath(__file__))
home_dir = os.path.dirname(current_dir)
template_path = os.path.join(home_dir, 'resources', 'Template_DailyWorkReport.xlsx')
output_path = os.path.join(current_dir, 'comparison_report.xlsx')

manager = DailyWorkReportManager(template_path)

data = {
    'date': datetime.date.today(),
    'vehicles': [
        {
            'out_locking': '잠김',
            'in_locking': '잠김',
            'out_exterior': '양호',
            'in_exterior': '양호',
            'out_cleanliness': '양호',
            'in_cleanliness': '양호',
            'out_cleaning': '함',
            'in_cleaning': '함',
            'vehicle_info': 'TEST-1234',
            'mileage': '100',
            'remarks': 'Comparison test'
        }
    ],
    'methods': {'RT': {}, 'UT': {}, 'MT': {}, 'PT': {}, 'PAUT': {}} # 5 methods -> 1 extra row
}

print("Generating report...")
manager.generate_report(data, output_path)

# Compare
wb_t = openpyxl.load_workbook(template_path)
ws_t = wb_t.active

wb_o = openpyxl.load_workbook(output_path)
ws_o = wb_o.active

def get_row_merges(ws, row):
    return [str(r) for r in ws.merged_cells.ranges if r.min_row <= row <= r.max_row]

def find_row_by_text(ws, text):
    for r in range(1, 100):
        for c in range(1, 5):
            val = str(ws.cell(row=r, column=c).value or "")
            if text in val:
                return r
    return None

def find_row_by_text_all_cols(ws, text):
    for r in range(1, 120):
        for c in range(1, 20):
            val = str(ws.cell(row=r, column=c).value or "")
            if text in val:
                return r
    return None

t_s3_title_row = find_row_by_text_all_cols(ws_t, "3. 차량관리")
o_s3_title_row = find_row_by_text_all_cols(ws_o, "3. 차량관리")

t_chk_header_row = find_row_by_text_all_cols(ws_t, "차량 외부상태")
o_chk_header_row = find_row_by_text_all_cols(ws_o, "차량 외부상태")

print(f"Template Section 3 Title Row: {t_s3_title_row}")
print(f"Output Section 3 Title Row: {o_s3_title_row}")
print(f"Template Merges on Row {t_s3_title_row}: {get_row_merges(ws_t, t_s3_title_row) if t_s3_title_row else 'N/A'}")
print(f"Output Merges on Row {o_s3_title_row}: {get_row_merges(ws_o, o_s3_title_row) if o_s3_title_row else 'N/A'}")

print(f"\nTemplate Checklist Header Row: {t_chk_header_row}")
print(f"Output Checklist Header Row: {o_chk_header_row}")
print(f"Template Merges on Row {t_chk_header_row}: {get_row_merges(ws_t, t_chk_header_row) if t_chk_header_row else 'N/A'}")
print(f"Output Merges on Row {o_chk_header_row}: {get_row_merges(ws_o, o_chk_header_row) if o_chk_header_row else 'N/A'}")
