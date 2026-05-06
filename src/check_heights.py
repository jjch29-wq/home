from openpyxl import load_workbook
import os

path = os.path.abspath("../resources/Template_DailyWorkReport.xlsx")
wb = load_workbook(path)
sheet = wb.active

titles = [12, 27, 41]
for r in titles:
    h = sheet.row_dimensions[r].height
    v = sheet.cell(row=r, column=1).value or sheet.cell(row=r, column=2).value
    print(f"Row {r} ({v}): Height={h}")

wb.close()
