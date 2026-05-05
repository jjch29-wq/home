from openpyxl import load_workbook
import os

path = os.path.abspath("../resources/Template_DailyWorkReport.xlsx")
wb = load_workbook(path)
sheet = wb.active

# Find Section 6 Header (usually row 42)
header_row = 42
for r in range(40, 50):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 20)]
    if any(x and "규격" in str(x) for x in row_vals):
        header_row = r
        print(f"Header Row {r}: {row_vals}")
        break

wb.close()
