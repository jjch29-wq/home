import sys
import os
import glob
import math
import traceback
import re
import warnings
import json
import tempfile
import datetime
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog

# --- Versioning ---
APP_VERSION = "v260404.01"
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, Border, Side
from PIL import Image as PILImage, ImageChops
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# Ignore library warnings
warnings.simplefilter("ignore")

# --- Constants & Paths ---
NAN_PATTERN = re.compile(r'^nan(\.0+)?$|^none$|^null$|^0\.0+|-0\.0+$', re.IGNORECASE)
DOT_ZERO_PATTERN = re.compile(r'\.0$')

# [PT] SCH -> 두께(mm) 변환 테이블
SCH_TO_THK = {
    "1/2": {"5S": 1.65, "10S": 2.11, "40": 2.77, "80": 3.73, "160": 4.75, "XXS": 7.47},
    "3/4": {"5S": 1.65, "10S": 2.11, "40": 2.87, "80": 3.91, "160": 5.56, "XXS": 7.82},
    "1": {"5S": 1.65, "10S": 2.77, "40": 3.38, "80": 4.55, "160": 6.35, "XXS": 9.09},
    "1-1/4": {"5S": 1.65, "10S": 2.77, "40": 3.56, "80": 4.85, "160": 6.35, "XXS": 9.70},
    "1-1/2": {"5S": 1.65, "10S": 2.77, "40": 3.68, "80": 5.08, "160": 7.14, "XXS": 10.16},
    "2": {"5S": 1.65, "10S": 2.77, "40": 3.91, "80": 5.54, "160": 8.74, "XXS": 11.07},
    "2-1/2": {"5S": 2.11, "10S": 3.05, "40": 5.16, "80": 7.01, "160": 9.53, "XXS": 14.02},
    "3": {"5S": 2.11, "10S": 3.05, "40": 5.49, "80": 7.62, "160": 11.13, "XXS": 15.24},
    "4": {"5S": 2.11, "10S": 3.05, "40": 6.02, "80": 8.56, "160": 13.49, "XXS": 17.12},
    "5": {"5S": 2.77, "10S": 3.40, "40": 6.55, "80": 9.53, "120": 12.70, "160": 15.88},
    "6": {"5S": 2.77, "10S": 3.40, "40": 7.11, "80": 10.97, "120": 14.27, "160": 18.26, "XXS": 21.95},
    "8": {"5S": 2.77, "10S": 3.76, "20": 6.35, "30": 7.04, "40": 8.18, "60": 10.31, "80": 12.70, "100": 15.09, "120": 18.26, "140": 20.62, "160": 23.01, "XXS": 22.23},
    "10": {"5S": 3.40, "10S": 4.19, "20": 6.35, "30": 7.80, "40": 9.27, "60": 12.70, "80": 15.09, "100": 18.26, "120": 21.44, "140": 25.40, "160": 28.58},
    "12": {"5S": 3.96, "10S": 4.57, "20": 6.35, "30": 8.38, "40": 10.31, "60": 14.27, "80": 17.48, "100": 21.44, "120": 25.40, "140": 28.58, "160": 33.32},
    "14": {"5S": 3.96, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 9.53, "40": 11.13, "60": 15.09, "80": 19.05, "100": 23.83, "120": 27.79, "140": 31.75, "160": 35.71},
    "16": {"5S": 4.19, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 9.53, "40": 12.70, "60": 16.66, "80": 21.44, "100": 26.19, "120": 30.96, "140": 36.53, "160": 40.49},
    "18": {"5S": 4.19, "10S": 4.78, "10": 6.35, "20": 7.92, "30": 11.13, "40": 14.27, "60": 19.05, "80": 23.83, "100": 29.36, "120": 34.93, "140": 39.67, "160": 45.24},
    "20": {"5S": 4.78, "10S": 5.54, "10": 6.35, "20": 9.53, "30": 12.70, "40": 15.09, "60": 20.62, "80": 26.19, "100": 32.54, "120": 38.10, "140": 44.45, "160": 50.01},
    "24": {"5S": 5.54, "10S": 6.35, "10": 6.35, "20": 9.53, "30": 14.27, "40": 17.48, "60": 24.61, "80": 30.96, "100": 38.89, "120": 46.02, "140": 52.37, "160": 59.54},
}

def convert_sch_to_thk(size_val, thk_val):
    """SCH 값을 두께(mm)로 변환"""
    if pd.isna(thk_val) or str(thk_val).strip() == "": return ""
    thk_str = str(thk_val).strip().upper()
    try:
        val = float(thk_str.replace("MM", "").replace("T", "").strip())
        if 0 < val < 100: return f"{val:.2f}"
    except: pass
    sch_match = re.search(r'(?:SCH[.\s]?|S/)?(\d+S?|XXS|XS)', thk_str, re.IGNORECASE)
    if not sch_match: return thk_str
    sch = sch_match.group(1).upper()
    if sch.endswith('S') and sch not in ['5S', '10S', 'XXS', 'XS']: sch = sch[:-1]
    if pd.isna(size_val) or str(size_val).strip() == "": return thk_str
    size_str = str(size_val).strip().replace('"', '').replace("'", "")
    size_str = re.sub(r'\s+', '-', size_str)
    if size_str in SCH_TO_THK and sch in SCH_TO_THK[size_str]:
        return f"{SCH_TO_THK[size_str][sch]:.2f}"
    try:
        size_int = str(int(float(size_str)))
        if size_int in SCH_TO_THK and sch in SCH_TO_THK[size_int]:
            return f"{SCH_TO_THK[size_int][sch]:.2f}"
    except: pass
    return thk_str

# 공통 스타일 (테두리 등)
thin_side = Side(style='thin')
medium_side = Side(style='medium')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

if getattr(sys, 'frozen', False):
    # PyInstaller creates a temp folder and stores path in _MEIPASS
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
    SCRIPT_HOME = base_path
    BASE_DIR = base_path
    CONFIG_DIR = base_path
    RESOURCE_DIR = os.path.join(base_path, "resources")
else:
    SCRIPT_HOME = os.path.dirname(os.path.abspath(__file__))
    BASE_DIR = os.path.dirname(SCRIPT_HOME)
    CONFIG_DIR = os.path.join(BASE_DIR, "config")
    RESOURCE_DIR = os.path.join(BASE_DIR, "resources")

SETTINGS_FILE = os.path.join(CONFIG_DIR, "logo_settings_unified.json")

class PMIReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"SITCO 통합 성적서 자동 생성기 ({APP_VERSION})")
        self.root.geometry("950x850") # Slightly wider/taller for multi-tab
        self.root.configure(background="#f9fafb")
        
        # 1. Initialize Configuration first (needed for state variables)
        self.config = {
            # UI State (Persistence)
            'PMI_SASH_RATIO': 0.5, 'RT_SASH_RATIO': 0.5, 'PT_SASH_RATIO': 0.5, 'PAUT_SASH_RATIO': 0.5,
            'PMI_COL_WIDTHS': {}, 'RT_COL_WIDTHS': {}, 'PT_COL_WIDTHS': {}, 'PAUT_COL_WIDTHS': {},
            # 갑지 (Cover)
            'SEOUL_COVER_ANCHOR': "E5", 'SEOUL_COVER_W': 200.0, 'SEOUL_COVER_H': 18.0, 'SEOUL_COVER_X': 30.0, 'SEOUL_COVER_Y': 15.0,
            'SITCO_COVER_ANCHOR': "A6", 'SITCO_COVER_W': 80.0, 'SITCO_COVER_H': 40.0, 'SITCO_COVER_X': 10.0, 'SITCO_COVER_Y': 10.0,
            'FOOTER_COVER_ANCHOR': "P50", 'FOOTER_COVER_W': 80.0, 'FOOTER_COVER_H': 20.0, 'FOOTER_COVER_X': 0.0, 'FOOTER_COVER_Y': 10.0,
            'FOOTER_PT_COVER_ANCHOR': "A50", 'FOOTER_PT_COVER_W': 100.0, 'FOOTER_PT_COVER_H': 25.0, 'FOOTER_PT_COVER_X': 5.0, 'FOOTER_PT_COVER_Y': 10.0,
            # 을지 (Data)
            'SEOUL_DATA_ANCHOR': "F5", 'SEOUL_DATA_W': 200.0, 'SEOUL_DATA_H': 18.0, 'SEOUL_DATA_X': 30.0, 'SEOUL_DATA_Y': 15.0,
            'SITCO_DATA_ANCHOR': "A6", 'SITCO_DATA_W': 100.0, 'SITCO_DATA_H': 50.0, 'SITCO_DATA_X': 20.0, 'SITCO_DATA_Y': 5.0,
            'FOOTER_DATA_ANCHOR': "M46", 'FOOTER_DATA_W': 100.0, 'FOOTER_DATA_H': 15.0, 'FOOTER_DATA_X': 25.0, 'FOOTER_DATA_Y': 10.0,
            'FOOTER_PT_DATA_ANCHOR': "A46", 'FOOTER_PT_DATA_W': 100.0, 'FOOTER_PT_DATA_H': 25.0, 'FOOTER_PT_DATA_X': 3.0, 'FOOTER_PT_DATA_Y': 7.0,
            'MARGIN_DATA_TOP': 0.2, 'MARGIN_DATA_BOTTOM': 0.0, 'MARGIN_DATA_LEFT': 0.5, 'MARGIN_DATA_RIGHT': 0.3, 'PRINT_SCALE_DATA': 95,
            'START_ROW': 19, 'DATA_END_ROW': 45, 'PRINT_END_ROW': 47,
            # [NEW] 선택적 행/열 조절용 설정
            'CUSTOM_ROWS_COVER': '23-38', 'CUSTOM_ROW_HEIGHT_COVER': 21.0,
            'CUSTOM_COLS_COVER': '', 'CUSTOM_COL_WIDTH_COVER': 10.0,
            'CUSTOM_ROWS_DATA': '', 'CUSTOM_ROW_HEIGHT_DATA': 20.55,
            'CUSTOM_COLS_DATA': '', 'CUSTOM_COL_WIDTH_DATA': 10.0,
            
            # PAUT Specific (B31.1)
            'PAUT_START_ROW': 11, 'PAUT_DATA_END_ROW': 40, 'PAUT_PRINT_END_ROW': 45,
            
            # RT Specific (Radiographic Testing)
            'RT_START_ROW': 11, 'RT_DATA_END_ROW': 45, 'RT_PRINT_END_ROW': 47
        }
        self.load_settings()

        # 2. State Variables
        self.logo_folder_path = tk.StringVar(value=RESOURCE_DIR)
        self.target_file_path = tk.StringVar(value=self.config.get('PMI_TARGET_PATH', ""))
        self.template_file_path = tk.StringVar(value=self.config.get('PMI_TEMPLATE_PATH', ""))
        self.sequence_filter = tk.StringVar()
        self.extraction_mode = tk.StringVar(value="전체")
        self.auto_verify = tk.BooleanVar(value=True)
        self.pmi_pane_ratio = self.config.get('PMI_SASH_RATIO', 0.5)
        
        self.show_selected_only = tk.BooleanVar(value=False)
        self.extracted_data = [] 
        self.pmi_sort_col = "" 
        self.pmi_sort_rev = False 
        self.rt_sort_col = ""
        self.rt_sort_rev = False
        self.pt_sort_col = ""
        self.pt_sort_rev = False
        self.paut_sort_col = ""
        self.paut_sort_rev = False
        
        # [NEW] PMI Preview Search & Filter
        self.pmi_search_loc = tk.StringVar()
        self.pmi_show_deficiency_only = tk.BooleanVar(value=False)
        
        # [NEW] 원소 함량 필터링용 상태 변수 (PMI 전용)
        self.element_filters = [] # list of dict: {'key': StringVar, 'min': StringVar, 'max': StringVar}
        # 기본 필터 추가 (Cr, Ni, Mo)
        for k in ["Cr", "Ni", "Mo"]:
            self.element_filters.append({
                'key': tk.StringVar(value=k), 
                'min': tk.StringVar(value=''), 
                'max': tk.StringVar(value='')
            })
            
        # --- PAUT State Variables ---
        self.paut_target_file_path = tk.StringVar(value=self.config.get('PAUT_TARGET_PATH', ""))
        self.paut_template_file_path = tk.StringVar(value=self.config.get('PAUT_TEMPLATE_PATH', ""))
        self.paut_manual_vars = {
            't': tk.StringVar(), 'h': tk.StringVar(), 'l': tk.StringVar(), 'd': tk.StringVar(),
            'nature': tk.StringVar(value="Slag"), 'loc': tk.StringVar(value="-"),
            'db': tk.StringVar(value="6"), 'peak': tk.StringVar(value="80"), 
            'z1': tk.StringVar(), 'z2': tk.StringVar(), 
            'L1': tk.StringVar(), 'L2': tk.StringVar(),
            'D_Upper': tk.StringVar(), 'D_Lower': tk.StringVar(),
            'target_fsh': tk.StringVar(value="40.0%")
        }
        self.paut_eval_mode = tk.StringVar(value="ECA") # [NEW] Standard / ECA
        self.paut_manual_vars['peak'].trace_add("write", lambda *args: self._update_paut_target_fsh())
        self.paut_manual_vars['db'].trace_add("write", lambda *args: self._update_paut_target_fsh())
        self.paut_extracted_data = []
        
        # --- RT State Variables ---
        self.rt_target_file_path = tk.StringVar(value=self.config.get('RT_TARGET_PATH', ""))
        self.rt_template_file_path = tk.StringVar(value=self.config.get('RT_TEMPLATE_PATH', ""))
        self.rt_extracted_data = []
        
        # --- PT State Variables ---
        self.pt_target_file_path = tk.StringVar(value=self.config.get('PT_TARGET_PATH', ""))
        self.pt_template_file_path = tk.StringVar(value=self.config.get('PT_TEMPLATE_PATH', ""))
        self.pt_extracted_data = []
        self.item_idx_map = []      # [NEW] Missing for PMI
        self.rt_item_idx_map = []   # [NEW] Missing for RT
        self.pt_item_idx_map = []
        self.paut_item_idx_map = [] 
        
        # [REFINED] Column Keys Mapping (Must match Treeview column count and order)
        self.column_keys = ["_status", "selected", "No", "Date", "Dwg", "Joint", "Loc", "Ni", "Cr", "Mo", "Grade"]
        self.rt_column_keys = ["selected", "No", "Date", "Dwg", "Joint", "Loc", "Acc", "Rej", "Deg", "D1", "D2", "D3", "D4", "D5", "D6", "D7", "D8", "D9", "D10", "D11", "D12", "D13", "D14", "D15", "Welder", "Remarks"]
        self.pt_column_keys = ["selected", "No", "Date", "Dwg", "Joint", "Material", "TestItem", "Result", "Welder", "Remarks"]
        self.paut_column_keys = ["selected", "No", "Line No.", "Joint No.", "Th'k(mm)", "Start", "End", "Length(mm)", "Upper", "Lower", "Height(mm)", "Type of Flaw", "a/l", "a/t", "Evaluation", "Remarks"]
        
        self.date_listbox = None
    
        # [NEW] Handle Application Closing for Final State Capture
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.rt_date_listbox = None
        self.pt_date_listbox = None
        self.paut_date_listbox = None

        # Keys are now verified to match Treeview column indices 1-N.

        # 3. UI Initialization
        # [NEW] File Info Strings for Preview Headers
        self.file_info_vars = {
            'PMI': tk.StringVar(value="📄 파일을 선택해주세요."),
            'RT': tk.StringVar(value="📄 파일을 선택해주세요."),
            'PT': tk.StringVar(value="📄 파일을 선택해주세요."),
            'PAUT': tk.StringVar(value="📄 파일을 선택해주세요.")
        }

        self.create_widgets()
        
        # [NEW] Sync initial file info from loaded config
        self._sync_all_file_infos()
        self.log("[INFO] 통합 버전을 시작했습니다.")
        
    def _safe_update_scrollregion(self):
        """Safely update canvas scrollregion, handling None bbox."""
        try:
            bbox = self.canvas.bbox("all")
            if bbox:
                self.canvas.configure(scrollregion=bbox)
        except Exception:
            pass

    def log(self, message):
        if hasattr(self, 'status_log'):
            self.status_log.config(state='normal')
            self.status_log.insert(tk.END, f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}\n")
            self.status_log.see(tk.END)
            self.status_log.config(state='disabled')
            self.root.update_idletasks()
        else:
            print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}")

    def fix_material_name(self, t):
        """재질명을 표준 포맷으로 변환"""
        t_str = str(t).upper()
        if t_str == 'NAN': return ""
        t_str = t_str.replace('A312-TP304', 'S/S').replace('A312-304L', 'S/S').replace('A312-305L', 'S/S').replace('A53-B', 'C/S').replace('A106-B', 'C/S')
        return t_str.replace('C2','C/S').replace('C4','C/S').replace('CS','C/S').replace('S99','S/S').replace('SS','S/S')

    def force_two_digit(self, val):
        """숫자 값을 2자리 문자열로 변환 (예: 1 -> 01)"""
        try:
            s = str(val).strip()
            f = float(s)
            if f.is_integer(): return f"{int(f):02d}"
            return s
        except: return str(val).strip()

    def load_settings(self):
        # 1. 내장 설정 파일 (실행파일 내부에 번들링된 기본값)
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            bundle_settings = os.path.join(sys._MEIPASS, "logo_settings_unified.json")
            if os.path.exists(bundle_settings):
                try:
                    with open(bundle_settings, 'r', encoding='utf-8') as f:
                        bundle_data = json.load(f)
                        self.config.update(bundle_data)
                    print("SUCCESS: 내장된 기본 설정을 로드했습니다.")
                except Exception as e:
                    print(f"WARNING: 내장 설정 불러오기 실패: {e}")

        # 2. 외부 설정 파일 (사용자 저장값 - 내장값보다 우선순위 높음)
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    saved_data = json.load(f)
                    
                    # [REFINED] Case-insensitive Sanitization for all modules
                    sanit_map = {
                        'PMI_NAME_DWG': (['ISO/DWG', 'ISO/DWG.'], 'Drawing No.'),
                        'PMI_NAME_JOINT': (['JOINT NO', 'JOINT NO.'], 'Joint No.'),
                        'PMI_NAME_LOC': (['TEST LOCATION', 'LOC'], 'Location'),
                        'PMI_NAME_RES': (['GRADE', 'RESULT'], 'Result'),
                        'RT_NAME_DWG': (['ISO/DWG', 'ISO/DWG.', 'DWG'], 'Drawing No.'),
                        'RT_NAME_JOINT': (['JOINT NO', 'FILM IDENT', 'JOINT'], 'Film Ident. No.'),
                        'RT_NAME_LOC': (['LOC', 'LOC.'], 'Film Location'),
                        'RT_NAME_WELDER': (['WELDER'], 'Welder No'),
                        'RT_NAME_DEG': (['물성', '물성(DEG)', 'DEG'], 'Deg')
                    }
                    for k, (olds, new) in sanit_map.items():
                        if k in saved_data:
                            val = str(saved_data[k]).upper().strip()
                            if any(val == o.upper() for o in olds):
                                saved_data[k] = new

                    # [NEW] Sanitize old "Element-" prefixes and Mn exclusion
                    for k in ['PMI_NAME_NI', 'PMI_NAME_CR', 'PMI_NAME_MO']:
                        if k in saved_data and isinstance(saved_data[k], str) and (saved_data[k].startswith("Element-") or "ELEMENT" in saved_data[k].upper()):
                            saved_data[k] = saved_data[k].replace("Element-", "").replace("ELEMENT-", "").replace("Element", "").strip()
                    
                    if 'column_keys' in saved_data and isinstance(saved_data['column_keys'], list):
                        while 'Mn' in saved_data['column_keys']: saved_data['column_keys'].remove('Mn')

                    # [NEW] Defect Names Sanitization (D1-D15 to circled labels)
                    defect_standards = {
                        1: "① Crack", 2: "② IP", 3: "③ LF", 4: "④ Slag", 5: "⑤ Por",
                        6: "⑥ U/C", 7: "⑦ RUC", 8: "⑧ BT", 9: "⑨ TI", 10: "⑩ CP",
                        11: "⑪ RC", 12: "⑫ Mis", 13: "⑬ EP", 14: "⑭ SD", 15: "⑮ Oth"
                    }
                    for d_i, std_name in defect_standards.items():
                        d_k = f'RT_NAME_D{d_i}'
                        if d_k in saved_data:
                            curr_val = str(saved_data[d_k]).upper().strip()
                            if curr_val == f"D{d_i}" or not curr_val or curr_val == "NAN":
                                saved_data[d_k] = std_name

                    self.config.update(saved_data)
                print("SUCCESS: 사용자 저장 설정을 적용했습니다.")
                
                # Restore Window Geometry/State
                if 'WINDOW_GEOMETRY' in saved_data:
                    self.root.geometry(saved_data['WINDOW_GEOMETRY'])
                if 'WINDOW_STATE' in saved_data:
                    try: self.root.state(saved_data['WINDOW_STATE'])
                    except: pass
                if 'ACTIVE_TAB' in saved_data:
                    self.root.after(200, lambda: self.mode_notebook.select(saved_data['ACTIVE_TAB']))
            except Exception as e:
                print(f"WARNING: 사용자 설정 불러오기 실패: {e}")

        # PAUT 마이그레이션 (UT_* -> PAUT_*)
        for k in list(self.config.keys()):
            if k.startswith("UT_"):
                new_k = k.replace("UT_", "PAUT_")
                if new_k not in self.config:
                    self.config[new_k] = self.config[k]

    def capture_ui_state(self):
        """UI 요소들의 현재 상태(컬럼 너비, 분할선 위치 등)를 config에 반영"""
        try:
            # 1. 트리뷰 컬럼 너비 캡처
            trees = [
                ('PMI', getattr(self, 'preview_tree', None)), 
                ('RT', getattr(self, 'rt_preview_tree', None)), 
                ('PT', getattr(self, 'pt_preview_tree', None)), 
                ('PAUT', getattr(self, 'paut_preview_tree', None))
            ]
            for mode, tree in trees:
                if tree:
                    try:
                        self.config[f"{mode}_COL_WIDTHS"] = {col: tree.column(col, "width") for col in tree["columns"]}
                    except: pass

            # 2. 패널 분할선(Sash) 비율 캡처
            paned_windows = [
                ('PMI', getattr(self, 'pmi_paned', None)),
                ('RT', getattr(self, 'rt_paned', None)),
                ('PT', getattr(self, 'pt_paned', None)),
                ('PAUT', getattr(self, 'paut_paned', None))
            ]
            for mode, pw in paned_windows:
                if pw:
                    try:
                        total_w = pw.winfo_width()
                        if total_w > 100:
                            self.config[f'{mode}_SASH_RATIO'] = pw.sash_coord(0)[0] / total_w
                    except: pass

            # 3. 윈도우 및 활성 탭 상태 캡처
            try:
                self.config['WINDOW_GEOMETRY'] = self.root.geometry()
                self.config['WINDOW_STATE'] = self.root.state()
                if hasattr(self, 'mode_notebook'):
                    self.config['ACTIVE_TAB'] = self.mode_notebook.index("current")
                
                # 4. 파일 경로 관성 저장
                self.config['PMI_TARGET_PATH'] = self.target_file_path.get()
                self.config['PMI_TEMPLATE_PATH'] = self.template_file_path.get()
                self.config['RT_TARGET_PATH'] = self.rt_target_file_path.get()
                self.config['RT_TEMPLATE_PATH'] = self.rt_template_file_path.get()
                self.config['PT_TARGET_PATH'] = self.pt_target_file_path.get()
                self.config['PT_TEMPLATE_PATH'] = self.pt_template_file_path.get()
                self.config['PAUT_TARGET_PATH'] = self.paut_target_file_path.get()
                self.config['PAUT_TEMPLATE_PATH'] = self.paut_template_file_path.get()
            except: pass
        except: pass

    def save_settings(self):
        """현재 설정을 파일(JSON)에 저장"""
        self.capture_ui_state()
        try:
            if hasattr(self, 'setting_vars'):
                for key, var in self.setting_vars.items():
                    val = var.get()
                    try:
                        if key.endswith(('_X', '_Y', '_W', '_H')) or 'MARGIN' in key:
                            self.config[key] = float(val)
                        elif 'SCALE' in key or key.endswith('_ROW'):
                            self.config[key] = int(float(val))
                        else:
                            self.config[key] = str(val)
                    except: pass

            # [NEW] 필터 설정 저장 (PMI)
            if hasattr(self, 'element_filters'):
                filter_data = []
                for f_item in self.element_filters:
                    filter_data.append({
                        'key': f_item['key'].get(),
                        'min': f_item['min'].get(),
                        'max': f_item['max'].get()
                    })
                self.config["PMI_FILTERS"] = filter_data

            # [NEW] 탭별 전용 행 설정 강제 저장
            self.config["PMI_START_ROW"] = self.config.get("START_ROW", 19)
            self.config["PMI_DATA_END_ROW"] = self.config.get("DATA_END_ROW", 45)
            self.config["PMI_PRINT_END_ROW"] = self.config.get("PRINT_END_ROW", 47)

            # [NEW] Window State Capture
            self.config['WINDOW_GEOMETRY'] = self.root.geometry()
            self.config['WINDOW_STATE'] = self.root.state()
            try: self.config['ACTIVE_TAB'] = self.mode_notebook.index("current")
            except: pass

            with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=4)
            self.log("[SUCCESS] 설정이 파일에 저장되었습니다.")
        except Exception as e:
            self.log(f"[WARNING] 설정 저장 실패: {e}")

    def on_closing(self):
        """애플리케이션 종료 시 최종 상태 저장"""
        try:
            self.capture_ui_state()
            self.save_settings()
        except: pass
        self.root.destroy()

    def _apply_sash_ratio(self, mode):
        """Helper to apply saved sash ratio with a slight delay to ensure UI mapping."""
        pw = getattr(self, f"{mode.lower()}_paned", None)
        ratio = self.config.get(f"{mode}_SASH_RATIO")
        if pw and ratio is not None:
            def apply():
                total_w = pw.winfo_width()
                if total_w > 100:
                    pw.sash_place(0, int(total_w * ratio), 0)
                else:
                    self.root.after(100, apply)
            self.root.after(100, apply)

    def evaluate_paut_flaw(self, t, h, l, depth, flaw_nature, mode="ECA"):
        """
        PAUT Acceptance Criteria Logic
        - Standard (ASME B31.1): Fatal flaws (Crack, LOF, IP) are rejected immediately.
        - ECA (Fracture Mechanics): Only Crack is fatal. LOF/IP are evaluated by size.
        """
        try:
            t_val = float(t)
            h_val = float(h)
            l_val = float(l)
            d_val = float(depth)
        except (ValueError, TypeError):
            return "Error (Invalid Dimension)", "Unknown"

        # 0. Automatic Location Determination
        s_top = d_val
        s_bottom = t_val - (d_val + h_val)
        s = min(s_top, s_bottom)
        s_limit = 0.4 * (h_val / 2)
        
        loc = "Surface" if s <= s_limit else "Subsurface"

        # 1. Immediate Rejection (Crack)
        nature_str = str(flaw_nature).strip().lower()
        if 'crack' in nature_str or '균열' in nature_str:
            return "Reject (Crack)", loc
            
        # [NEW] Mode-based Rejection for LOF (LF) and IP
        if mode == "Standard":
            unacceptable_types = ['lof', 'lack of fusion', 'ip', 'incomplete penetration', 'lf']
            if any(x in nature_str for x in unacceptable_types):
                return f"Reject ({flaw_nature})", loc
        
        if l_val <= 0 or h_val <= 0 or t_val <= 0:
            return "Error (Zero/Negative Value)", loc

        # 1.1 Special Rules for 6mm <= t < 13mm
        if 6 <= t_val < 13:
            if l_val > 6.4: return f"Reject (L: {l_val} > 6.4mm)", loc
            if t_val < 10: h_surf_max, h_sub_max = 0.95, 0.96
            elif t_val < 12: h_surf_max, h_sub_max = 1.04, 1.04
            else: h_surf_max, h_sub_max = 1.13, 1.14
                
            limit = h_surf_max if loc == "Surface" else h_sub_max
            if h_val > limit: return f"Reject ({loc} h: {h_val} > {limit}mm)", loc
            return "Accept", loc

        # 1.2 Special Rules for 13mm <= t < 25.4mm
        if 13 <= t_val < 25.4:
            if l_val > 6.4: return f"Reject (L: {l_val} > 6.4mm)", loc
            actual_h_t = h_val / t_val
            allowed_h_t = 0.087 if loc == "Surface" else 0.143 
            if actual_h_t > allowed_h_t: return f"Reject ({loc} h/t: {actual_h_t:.3f} > {allowed_h_t:.3f})", loc
            return "Accept", loc
                
        # 2. Aspect Ratio (a/l) logic for t >= 25.4mm
        a_val = h_val if loc == "Surface" else h_val / 2
        aspect_ratio_a_l = a_val / l_val
        
        master_table = [
            (0.00, 0.031, 0.034), (0.05, 0.033, 0.038), (0.10, 0.036, 0.043),
            (0.15, 0.041, 0.054), (0.20, 0.047, 0.066), (0.25, 0.055, 0.078),
            (0.30, 0.064, 0.090), (0.35, 0.074, 0.103), (0.40, 0.083, 0.116),
            (0.45, 0.085, 0.129), (0.50, 0.087, 0.143)
        ]
        
        allowed_a_t = 0
        for ar_limit, surf_a_t, sub_a_t in master_table:
            if aspect_ratio_a_l <= ar_limit:
                allowed_a_t = surf_a_t if loc == 'Surface' else sub_a_t
                break
        if allowed_a_t == 0: allowed_a_t = master_table[-1][1] if loc == 'Surface' else master_table[-1][2]

        actual_a_t = a_val / t_val
        if actual_a_t <= allowed_a_t: return "Accept", loc
        return f"Reject ({loc} a/t: {actual_a_t:.3f} > {allowed_a_t:.3f})", loc

    def create_widgets(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TFrame", background="#f9fafb")
        style.configure("TLabel", background="#f9fafb", font=("Malgun Gothic", 10))
        style.configure("TLabelframe", background="#f9fafb", font=("Malgun Gothic", 10, "bold"))
        style.configure("TLabelframe.Label", background="#f9fafb", font=("Malgun Gothic", 10, "bold"))
        style.configure("Action.TButton", font=("Malgun Gothic", 11, "bold"), padding=10)
        
        # [NEW] 탭 스타일 설정
        style.configure("Main.TNotebook.Tab", font=("Malgun Gothic", 11, "bold"), padding=[10, 3])
        
        style.map("TEntry", 
                  selectbackground=[('focus', '#3b82f6'), ('!focus', '#3b82f6')],
                  selectforeground=[('focus', 'white'), ('!focus', 'white')])

        # --- Main Scrollable Container ---
        self.canvas = tk.Canvas(self.root, background="#f9fafb", highlightthickness=0, yscrollincrement=40)
        self.scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, background="#f9fafb")

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self._safe_update_scrollregion()
        )

        self.canvas_frame_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        def _on_canvas_configure(event):
            # [ROOT CAUSE FIX] Always force width, handle None bbox gracefully
            self.canvas.itemconfigure(self.canvas_frame_window, width=event.width)
            self._safe_update_scrollregion()

        self.canvas.bind("<Configure>", _on_canvas_configure)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        self.root.option_add('*selectBackground', '#3b82f6')
        self.root.option_add('*selectForeground', 'white')

        # Treeview Style
        style.configure("Treeview", rowheight=28, font=("Malgun Gothic", 10))
        style.map("Treeview", background=[('selected', '#3b82f6')], foreground=[('selected', 'white')])
        style.configure("Treeview.Heading", font=("Malgun Gothic", 10, "bold"))

        self._create_entry_context_menu()
        self._create_tree_context_menu()

        # Mouse Wheel Support
        def _on_mousewheel(event):
            try:
                focused = self.root.focus_get()
                curr = focused
                while curr:
                    try:
                        c_name = curr.winfo_class()
                        if c_name in ('Treeview', 'Text', 'Listbox'):
                            curr.yview_scroll(int(-1*(event.delta/120)*4.0), "units")
                            return "break"
                    except: pass
                    curr = curr.master
            except Exception: pass
            self.canvas.yview_scroll(int(-1*(event.delta/120)*2.0), "units")
            return "break"
        self.root.bind_all("<MouseWheel>", _on_mousewheel, add="+")

        def _on_root_click(event):
            try:
                if "treeview" in str(event.widget).lower(): return
                w_class = event.widget.winfo_class()
                if w_class in ('Frame', 'TFrame', 'Label', 'TLabel', 'Canvas', 'Labelframe', 'TLabelframe'):
                    self.root.focus_set()
            except: pass
        self.root.bind_all("<Button-1>", _on_root_click, add="+")

        # [ROOT CAUSE FIX] Place Notebook DIRECTLY in root, NOT in Canvas scrollable_frame.
        # Canvas windows do not reliably propagate width to children.
        # Since PMI uses splitscreen (no scrolling needed), direct packing guarantees full width.
        
        # Hide Canvas & Scrollbar (no longer needed for layout)
        self.canvas.pack_forget()
        self.scrollbar.pack_forget()
        
        # ===== PACK ORDER: bottom-most items first =====
        
        # [UX] ESC Key Binding: Unfocus only (preserving content)
        self.root.bind_class("TEntry", "<Escape>", self._on_entry_esc)
        self.root.bind_class("Entry", "<Escape>", self._on_entry_esc)
        
        # 1. Status bar (very bottom)
        self.status_bar = tk.Frame(self.root, background="#e2e8f0", height=22)
        self.status_bar.pack(fill='x', side='bottom')
        self.status_bar.pack_propagate(False)
        tk.Label(self.status_bar, text="준비됨", font=("Malgun Gothic", 8), background="#e2e8f0", foreground="#64748b").pack(side='left', padx=10)
        tk.Label(self.status_bar, text=f"Build Version: v{APP_VERSION}", font=("Arial", 8), background="#e2e8f0", foreground="#94a3b8").pack(side='right', padx=10)

        # 2. Progress bar + Log (above status bar, compact)
        bottom_frame = tk.Frame(self.root, background="#f9fafb")
        bottom_frame.pack(fill='x', side='bottom', padx=5)

        self.progress = ttk.Progressbar(bottom_frame, orient="horizontal", mode="determinate")
        self.progress.pack(fill='x', pady=(5, 2))

        log_frame = tk.Frame(bottom_frame, background="#f9fafb")
        log_frame.pack(fill='x')

        self.status_log = tk.Text(log_frame, height=4, font=("Consolas", 8), state='disabled', background="#1e1e2e", foreground="#10b981", padx=5, pady=3)
        vsb = ttk.Scrollbar(log_frame, orient="vertical", command=self.status_log.yview)
        self.status_log.configure(yscrollcommand=vsb.set)
        self.status_log.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        
        # 3. THEN pack Notebook (fills all remaining space)
        self.mode_notebook = ttk.Notebook(self.root, style="Main.TNotebook")
        self.mode_notebook.pack(fill='both', expand=True, padx=5, pady=(5, 0))
        
        self.pmi_mode_frame = tk.Frame(self.mode_notebook, background="#f9fafb")
        self.rt_mode_frame = tk.Frame(self.mode_notebook, background="#f9fafb")
        self.pt_mode_frame = tk.Frame(self.mode_notebook, background="#f9fafb")
        self.paut_mode_frame = tk.Frame(self.mode_notebook, background="#f9fafb")
        
        self.mode_notebook.add(self.pmi_mode_frame, text="  PMI (성분 분석)  ", sticky='nsew')
        self.mode_notebook.add(self.rt_mode_frame, text="  RT (방사선 투과)  ", sticky='nsew')
        self.mode_notebook.add(self.pt_mode_frame, text="  PT (침투 탐상)  ", sticky='nsew')
        self.mode_notebook.add(self.paut_mode_frame, text="  PAUT (ASME B31.1)  ", sticky='nsew')
        
        # Setup each mode
        self._setup_pmi_ui(self.pmi_mode_frame)
        self._setup_rt_ui(self.rt_mode_frame)
        self._setup_pt_ui(self.pt_mode_frame)
        self._setup_paut_ui(self.paut_mode_frame)

    def _create_scrollable_sidebar(self, parent):
        """Creates a scrollable canvas/scrollbar container for sidebars."""
        canvas = tk.Canvas(parent, background="#f9fafb", highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, background="#f9fafb", padx=10, pady=10)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        def _on_canvas_configure(e):
            canvas.itemconfig(canvas_window, width=e.width)
        canvas.bind("<Configure>", _on_canvas_configure)
        
        canvas.configure(yscrollcommand=scrollbar.set)
        
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            return "break"
        
        canvas.bind('<Enter>', lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
        canvas.bind('<Leave>', lambda e: canvas.unbind_all("<MouseWheel>"))

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        return scrollable_frame

    def _setup_pmi_ui(self, parent):
        # [FORCE] Ensure parent (pmi_mode_frame) allows expansion
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # --- Dual Pane Layout (PanedWindow) ---
        self.pmi_paned = tk.PanedWindow(container, orient='horizontal', background="#d1d5db", 
                            sashwidth=6, sashpad=0, sashrelief='raised', borderwidth=0)
        self.pmi_paned.pack(fill='both', expand=True)

        # [LEFT] Filter/Settings Sidebar
        left_pane = tk.Frame(self.pmi_paned, background="#f9fafb", padx=10, pady=10)
        self.pmi_paned.add(left_pane, stretch="always")

        # Bind events to maintain ratio
        self.pmi_paned.bind("<Configure>", lambda e: [self._on_pmi_paned_configure(e), self.root.update_idletasks()])
        self.pmi_paned.bind("<ButtonRelease-1>", lambda e: self.root.after(10, self._update_pmi_ratio))

        # Header
        header_frame = tk.Frame(left_pane, background="#f9fafb")
        header_frame.pack(fill='x', pady=(0, 10))
        tk.Label(header_frame, text="🔬 PMI 관리", font=("Malgun Gothic", 15, "bold"), 
                 background="#f9fafb", foreground="#1e3a8a").pack(side='left')
        tk.Label(header_frame, text=f"v{APP_VERSION}", font=("Arial", 8), 
                 background="#f1f5f9", foreground="#64748b", padx=5, pady=1).pack(side='left', padx=10)

        # 1. File Selection Group
        file_container = ttk.LabelFrame(left_pane, text=" 데이터 및 양식 (Data) ", padding=10)
        file_container.pack(fill='x', pady=(0, 10))

        def _add_very_compact_row(parent_f, label, var, row, is_dir=False, types=None):
            parent_f.columnconfigure(1, weight=1)
            ttk.Label(parent_f, text=label, font=("Malgun Gothic", 8)).grid(row=row, column=0, sticky='e', padx=2, pady=2)
            # Use width=1 for entry to allow proactive shrinking
            ttk.Entry(parent_f, textvariable=var, font=("Arial", 9), width=1, exportselection=False).grid(row=row, column=1, padx=2, pady=2, sticky='ew')
            cmd = (lambda : self._browse_dir(var)) if is_dir else (lambda : self._browse_file(var, types))
            ttk.Button(parent_f, text="...", width=3, command=cmd).grid(row=row, column=2, padx=2, pady=2)

        _add_very_compact_row(file_container, "로고:", self.logo_folder_path, 0, is_dir=True)
        _add_very_compact_row(file_container, "데이터:", self.target_file_path, 1, types=[("Excel Source", "*.xls;*.xlsx;*.xlsm")])
        _add_very_compact_row(file_container, "양식:", self.template_file_path, 2, types=[("Excel Template", "*.xlsx;*.xlsm")])

        # 2. Configuration Notebook (Cover, Data, Rows - NO Preview tab here)
        config_frame = ttk.LabelFrame(left_pane, text=" 설정 (Config) ", padding=2)
        config_frame.pack(fill='both', expand=True, pady=(0, 10))

        self.tab_notebook = ttk.Notebook(config_frame)
        self.tab_notebook.pack(fill='both', expand=True)

        tab_cover = ttk.Frame(self.tab_notebook, padding=5)
        tab_data = ttk.Frame(self.tab_notebook, padding=5)
        tab_rows = ttk.Frame(self.tab_notebook, padding=5)
        tab_filter = ttk.Frame(self.tab_notebook, padding=10)
        tab_cols = ttk.Frame(self.tab_notebook, padding=5)
        
        # [CRITICAL] Allow children to expand within tabs
        for t in [tab_cover, tab_data, tab_rows, tab_filter, tab_cols]: t.columnconfigure(0, weight=1)

        self.tab_notebook.add(tab_cover, text="갑지")
        self.tab_notebook.add(tab_data, text="을지")
        self.tab_notebook.add(tab_rows, text="행 설정")
        self.tab_notebook.add(tab_cols, text="컬럼 설정")
        self.tab_notebook.add(tab_filter, text="필터/옵션")

        self.setting_vars = {}
        # [ALIGNED] Direct grid expansion (as in RT/PT)
        next_row_cover = self._create_setting_grid(tab_cover, "COVER")
        next_row_data = self._create_setting_grid(tab_data, "DATA")
        self._create_margin_settings(tab_cover, "COVER", next_row_cover)
        self._create_margin_settings(tab_data, "DATA", next_row_data)
        self._create_row_settings(tab_rows, mode="PMI")
        
        pmi_items = [
            ("No:", "PMI_COL_NO", 13, "PMI_NAME_NO", "No", "No"),
            ("Drawing No:", "PMI_COL_DWG", 1, "PMI_NAME_DWG", "Drawing No.", "Drawing No."),
            ("Joint No:", "PMI_COL_JOINT", 6, "PMI_NAME_JOINT", "Joint No.", "Joint No."),
            ("Location:", "PMI_COL_LOC", 7, "PMI_NAME_LOC", "Location", "Location"),
            ("Ni:", "PMI_COL_NI", 8, "PMI_NAME_NI", "Ni", "Ni"),
            ("Cr:", "PMI_COL_CR", 9, "PMI_NAME_CR", "Cr", "Cr"),
            ("Mo:", "PMI_COL_MO", 10, "PMI_NAME_MO", "Mo", "Mo"),
            ("판정 결과:", "PMI_COL_RES", 13, "PMI_NAME_RES", "Result", "Result")
        ]
        self._create_column_mapping_ui(tab_cols, "PMI", pmi_items)
        
        # Spacer for tabs
        tk.Frame(tab_rows, background="#f9fafb").pack(fill='x')

        # 3. Quick Options Section (Moved to Settings Notebook)
        # Filter Tag Area (Self-updating)
        filter_box = tk.Frame(tab_filter, background="#f9fafb")
        filter_box.pack(fill='x', pady=2)
        ttk.Label(filter_box, text="🔍 성분 필터:", background="#f9fafb", font=("Malgun Gothic", 9, "bold")).pack(side='left', padx=(0, 5))
        
        def add_filter():
            self.element_filters.append({'key': tk.StringVar(value=''), 'min': tk.StringVar(value=''), 'max': tk.StringVar(value='')})
            self._update_pmi_filter_ui()
        ttk.Button(filter_box, text="+", width=2, command=add_filter).pack(side='right')

        self.filter_container = tk.Frame(filter_box, background="#f9fafb")
        self.filter_container.pack(side='left', fill='x', expand=True)

        # Sequence Filter & Auto Check
        tk.Frame(tab_filter, height=1, background="#e5e7eb").pack(fill='x', pady=10)
        tk.Checkbutton(tab_filter, text="✅ 재질 자동 판정", variable=self.auto_verify, background="#f9fafb", font=("Malgun Gothic", 8)).pack(anchor='w', pady=(0, 5))
        seq_row = tk.Frame(tab_filter, background="#f9fafb")
        seq_row.pack(fill='x', pady=2)
        ttk.Label(seq_row, text="📊 특정순번:", background="#f9fafb", font=("Malgun Gothic", 8)).pack(side='left')
        ttk.Entry(seq_row, textvariable=self.sequence_filter, width=15, font=("Arial", 9), exportselection=False).pack(side='left', padx=5, fill='x', expand=True)

        action_bar = tk.Frame(left_pane, background="#f9fafb")
        action_bar.pack(fill='x', pady=5)
        ttk.Button(action_bar, text=" ✨ 생성 시작 ", style="Action.TButton", command=self.run_process).pack(fill='x', pady=(0, 5))
        ttk.Button(action_bar, text=" 📝 데이터 추출 ", command=self.extract_only).pack(fill='x')

        self._update_pmi_filter_ui()

        # [RIGHT] Data Preview Pane (Always visible)
        right_frame = ttk.LabelFrame(self.pmi_paned, text=" 🔬 실시간 데이터 관리 (Live Preview) ", padding=10)
        self.pmi_paned.add(right_frame, stretch="always")
        self._create_preview_ui(right_frame)
        self._apply_sash_ratio("PMI")
        

    def _update_pmi_ratio(self):
        """Saves current sash position and persists it."""
        try:
            total_w = self.pmi_paned.winfo_width()
            if total_w > 100:
                current_sash = self.pmi_paned.sash_coord(0)[0]
                self.pmi_pane_ratio = current_sash / total_w
                self.config['PMI_SASH_RATIO'] = self.pmi_pane_ratio
                self.save_settings() # Persist immediately
        except: pass

    def _on_pmi_paned_configure(self, event):
        """Maintains the proportional split when the window resizes."""
        try:
            total_w = self.pmi_paned.winfo_width()
            if total_w > 100:
                # During window resize, RE-PLACE sash based on SAVED ratio
                if event and event.widget == self.pmi_paned:
                    new_pos = int(total_w * self.pmi_pane_ratio)
                    self.pmi_paned.sash_place(0, new_pos, 0)
        except: pass

    def _update_pmi_filter_ui(self):
        """Redraws the element filter tags in a modern, elastic style."""
        for widget in self.filter_container.winfo_children():
            widget.destroy()
        
        def remove_filter(idx):
            self.element_filters.pop(idx)
            self._update_pmi_filter_ui()

        # [HYPER-REACTIVE] Use grid for even horizontal distribution (Elastic, Stacked)
        cols_per_row = 1
        for i, f_item in enumerate(self.element_filters):
            self.filter_container.columnconfigure(0, weight=1)
            
            # Create an "Elastic Tag" 
            tag = tk.Frame(self.filter_container, background="#eff6ff", highlightthickness=1, 
                            highlightbackground="#bfdbfe", padx=2, pady=1)
            tag.grid(row=i, column=0, sticky='ew', padx=2, pady=2)
            
            # Inner grid for tag components
            tag.columnconfigure(0, weight=1, minsize=0) # Key
            tag.columnconfigure(2, weight=1, minsize=0) # Min
            tag.columnconfigure(4, weight=1, minsize=0) # Max
            
            # Key
            e_key = tk.Entry(tag, textvariable=f_item['key'], width=1, font=("Arial", 8), 
                             relief='flat', background="#eff6ff", justify='center')
            e_key.grid(row=0, column=0, sticky='ew')
            tk.Label(tag, text=":", background="#eff6ff", font=("Arial", 8)).grid(row=0, column=1)
            
            # Min
            e_min = tk.Entry(tag, textvariable=f_item['min'], width=1, font=("Arial", 8), 
                             relief='flat', background="#eff6ff", justify='center')
            e_min.grid(row=0, column=2, sticky='ew')
            tk.Label(tag, text="~", background="#eff6ff", font=("Arial", 8)).grid(row=0, column=3)
            
            # Max
            e_max = tk.Entry(tag, textvariable=f_item['max'], width=1, font=("Arial", 8), 
                             relief='flat', background="#eff6ff", justify='center')
            e_max.grid(row=0, column=4, sticky='ew')
            
            # Delete button (Small)
            btn_del = tk.Button(tag, text="×", command=lambda idx=i: remove_filter(idx), 
                                relief='flat', background="#eff6ff", foreground="#3b82f6", font=("Arial", 9, "bold"),
                                activebackground="#dbeafe", cursor="hand2", padx=0, pady=0)
            btn_del.grid(row=0, column=5, padx=(2, 0))

    def _setup_rt_ui(self, parent):
        # [FORCE] Ensure parent (rt_mode_frame) allows expansion
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # --- Dual Pane Layout (PanedWindow) ---
        self.rt_paned = tk.PanedWindow(container, orient='horizontal', background="#d1d5db", 
                            sashwidth=6, sashpad=0, sashrelief='raised', borderwidth=0)
        self.rt_paned.pack(fill='both', expand=True)

        # [LEFT] Scrollable Settings Sidebar
        left_container = tk.Frame(self.rt_paned, background="#f9fafb")
        self.rt_paned.add(left_container, stretch="always")
        
        left_pane = self._create_scrollable_sidebar(left_container)

        # Header
        header_frame = tk.Frame(left_pane, background="#f9fafb")
        header_frame.pack(fill='x', pady=(0, 10))
        tk.Label(header_frame, text="🔬 RT 성적서 관리", font=("Malgun Gothic", 15, "bold"), 
                 background="#f9fafb", foreground="#1e3a8a").pack(side='left')

        # 1. File Selection Group
        file_frame = ttk.LabelFrame(left_pane, text=" 데이터 및 양식 (Data) ", padding=10)
        file_frame.pack(fill='x', pady=(0, 10))

        def _add_file_row(parent_frame, label, var, row, is_dir=False, types=None):
            parent_frame.columnconfigure(1, weight=1)
            ttk.Label(parent_frame, text=label, font=("Malgun Gothic", 8)).grid(row=row, column=0, sticky='e', padx=2, pady=2)
            ttk.Entry(parent_frame, textvariable=var, font=("Arial", 9), width=1, exportselection=False).grid(row=row, column=1, padx=2, pady=2, sticky='ew')
            cmd = (lambda: self._browse_dir(var)) if is_dir else (lambda: self._browse_file(var, types))
            ttk.Button(parent_frame, text="...", width=3, command=cmd).grid(row=row, column=2, padx=2, pady=2)

        _add_file_row(file_frame, "로고 폴더:", self.logo_folder_path, 0, is_dir=True)
        _add_file_row(file_frame, "RT 데이터:", self.rt_target_file_path, 1, types=[("Excel Source", "*.xls;*.xlsx;*.xlsm")])
        _add_file_row(file_frame, "RT 양식:", self.rt_template_file_path, 2, types=[("Excel Template", "*.xlsx;*.xlsm")])

        # 2. Configuration Tabs
        rt_config_frame = ttk.LabelFrame(left_pane, text=" 리포트 세부 설정 ", padding=2)
        rt_config_frame.pack(fill='both', expand=True, pady=(0, 10))

        self.rt_tab_notebook = ttk.Notebook(rt_config_frame)
        self.rt_tab_notebook.pack(fill='both', expand=True)

        rt_tab_cover = ttk.Frame(self.rt_tab_notebook, padding=5)
        rt_tab_data = ttk.Frame(self.rt_tab_notebook, padding=5)
        rt_tab_rows = ttk.Frame(self.rt_tab_notebook, padding=5)
        rt_tab_cols = ttk.Frame(self.rt_tab_notebook, padding=5)
        
        # [CRITICAL] Allow children to expand
        for t in [rt_tab_cover, rt_tab_data, rt_tab_rows, rt_tab_cols]: t.columnconfigure(0, weight=1)

        self.rt_tab_notebook.add(rt_tab_cover, text="갑지")
        self.rt_tab_notebook.add(rt_tab_data, text="을지")
        self.rt_tab_notebook.add(rt_tab_rows, text="행 설정")
        self.rt_tab_notebook.add(rt_tab_cols, text="컬럼 설정")

        next_row_rt_cover = self._create_setting_grid(rt_tab_cover, "COVER")
        next_row_rt_data = self._create_setting_grid(rt_tab_data, "DATA")
        self._create_margin_settings(rt_tab_cover, "COVER", next_row_rt_cover)
        self._create_margin_settings(rt_tab_data, "DATA", next_row_rt_data)
        self._create_row_settings(rt_tab_rows, mode="RT")
        
        rt_items = [
            ("No:", "RT_COL_NO", 1, "RT_NAME_NO", "No", "No"),
            ("Date:", "RT_COL_DATE", 2, "RT_NAME_DATE", "Date", "Date"),
            ("Drawing No.:", "RT_COL_DWG", 3, "RT_NAME_DWG", "Drawing No.", "Drawing No."),
            ("Film Ident. No.:", "RT_COL_JOINT", 4, "RT_NAME_JOINT", "Film Ident. No.", "Film Ident. No."),
            ("Film Location:", "RT_COL_LOC", 5, "RT_NAME_LOC", "Film Location", "Film Location"),
            ("두께(T):", "RT_COL_THK", 6, "RT_NAME_THK", "T", "T"),
            ("재질(Mat):", "RT_COL_MAT", 7, "RT_NAME_MAT", "Mat", "Mat"),
            ("물성(Deg):", "RT_COL_DEG", 10, "RT_NAME_DEG", "Deg", "Deg"),
            ("Acc:", "RT_COL_ACC", 8, "RT_NAME_ACC", "Acc", "Acc"),
            ("Rej:", "RT_COL_REJ", 9, "RT_NAME_REJ", "Rej", "Rej"),
            ("① Crack:", "RT_COL_D1", 13, "RT_NAME_D1", "① Crack", "① Crack"),
            ("② IP:", "RT_COL_D2", 14, "RT_NAME_D2", "② IP", "② IP"),
            ("③ LF:", "RT_COL_D3", 15, "RT_NAME_D3", "③ LF", "③ LF"),
            ("④ Slag:", "RT_COL_D4", 16, "RT_NAME_D4", "④ Slag", "④ Slag"),
            ("⑤ Por:", "RT_COL_D5", 17, "RT_NAME_D5", "⑤ Por", "⑤ Por"),
            ("⑥ U/C:", "RT_COL_D6", 18, "RT_NAME_D6", "⑥ U/C", "⑥ U/C"),
            ("⑦ RUC:", "RT_COL_D7", 19, "RT_NAME_D7", "⑦ RUC", "⑦ RUC"),
            ("⑧ BT:", "RT_COL_D8", 20, "RT_NAME_D8", "⑧ BT", "⑧ BT"),
            ("⑨ TI:", "RT_COL_D9", 21, "RT_NAME_D9", "⑨ TI", "⑨ TI"),
            ("⑩ CP:", "RT_COL_D10", 22, "RT_NAME_D10", "⑩ CP", "⑩ CP"),
            ("⑪ RC:", "RT_COL_D11", 23, "RT_NAME_D11", "⑪ RC", "⑪ RC"),
            ("⑫ Mis:", "RT_COL_D12", 24, "RT_NAME_D12", "⑫ Mis", "⑫ Mis"),
            ("⑬ EP:", "RT_COL_D13", 25, "RT_NAME_D13", "⑬ EP", "⑬ EP"),
            ("⑭ SD:", "RT_COL_D14", 26, "RT_NAME_D14", "⑭ SD", "⑭ SD"),
            ("⑮ Oth:", "RT_COL_D15", 27, "RT_NAME_D15", "⑮ Oth", "⑮ Oth"),
            ("판정(Result):", "RT_COL_RES", 28, "RT_NAME_RES", "Result", "Result"),
            ("용접사(Welder):", "RT_COL_WELDER", 29, "RT_NAME_WELDER", "Welder No", "Welder No"),
            ("비고(Remarks):", "RT_COL_REM", 30, "RT_NAME_REM", "Remarks", "Remarks")
        ]
        self._create_column_mapping_ui(rt_tab_cols, "RT", rt_items)

        # 3. Action Section
        action_frame = tk.Frame(left_pane, background="#ffffff", highlightthickness=1, highlightbackground="#d1d5db", padx=10, pady=5)
        action_frame.pack(fill='x', pady=(0, 10))

        filter_row = tk.Frame(action_frame, background="#ffffff")
        filter_row.pack(fill='x', pady=2)
        ttk.Label(filter_row, text="📊 특정순번:", background="#ffffff", font=("Malgun Gothic", 8)).pack(side='left')
        ttk.Entry(filter_row, textvariable=self.sequence_filter, width=15, font=("Arial", 9), exportselection=False).pack(side='left', padx=5, fill='x', expand=True)

        btn_row = tk.Frame(left_pane, background="#f9fafb")
        btn_row.pack(fill='x', pady=5)
        ttk.Button(btn_row, text=" ✨ 성적서 생성 ", style="Action.TButton", command=self.run_process).pack(fill='x', pady=(0, 5))
        ttk.Button(btn_row, text=" 📝 데이터 추출 ", command=self.extract_only).pack(fill='x')

        # [RIGHT] Preview Pane
        right_frame = ttk.LabelFrame(self.rt_paned, text=" 🔬 실시간 데이터 관리 (RT Preview) ", padding=10)
        self.rt_paned.add(right_frame, stretch="always")
        self._create_rt_preview_ui(right_frame)
        self._apply_sash_ratio("RT")
        

        # Adaptive Resizing Bindings
        self.rt_pane_ratio = self.config.get('RT_SASH_RATIO', 0.5)
        self.rt_paned.bind("<Configure>", lambda e: [self._on_rt_paned_configure(e), self.root.update_idletasks()])
        self.rt_paned.bind("<ButtonRelease-1>", lambda e: self.root.after(10, self._update_rt_ratio))
        self.root.after(500, lambda: self._on_rt_paned_configure(None))

    def _update_rt_ratio(self):
        try:
            total_w = self.rt_paned.winfo_width()
            if total_w > 100:
                current_sash = self.rt_paned.sash_coord(0)[0]
                self.rt_pane_ratio = current_sash / total_w
                self.config['RT_SASH_RATIO'] = self.rt_pane_ratio
                self.save_settings()
        except: pass

    def _on_rt_paned_configure(self, event):
        try:
            total_w = self.rt_paned.winfo_width()
            if total_w > 100:
                if event and event.widget == self.rt_paned:
                    new_pos = int(total_w * self.rt_pane_ratio)
                    self.rt_paned.sash_place(0, new_pos, 0)
        except: pass

    def _setup_pt_ui(self, parent):
        # [FORCE] Ensure parent (pt_mode_frame) allows expansion
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # --- Dual Pane Layout (PanedWindow) ---
        self.pt_paned = tk.PanedWindow(container, orient='horizontal', background="#d1d5db", 
                            sashwidth=6, sashpad=0, sashrelief='raised', borderwidth=0)
        self.pt_paned.pack(fill='both', expand=True)

        # [LEFT] Evaluation Controls
        left_pane = tk.Frame(self.pt_paned, background="#f9fafb", padx=10, pady=10)
        self.pt_paned.add(left_pane, stretch="always")

        # Header
        header_frame = tk.Frame(left_pane, background="#f9fafb")
        header_frame.pack(fill='x', pady=(0, 10))
        tk.Label(header_frame, text="🔬 PT 성적서 관리", font=("Malgun Gothic", 15, "bold"), 
                 background="#f9fafb", foreground="#1e3a8a").pack(side='left')

        # 1. File Selection Group
        file_frame = ttk.LabelFrame(left_pane, text=" 데이터 및 양식 (Data) ", padding=10)
        file_frame.pack(fill='x', pady=(0, 10))

        def _add_file_row(parent_f, label, var, row, is_dir=False, types=None):
            parent_f.columnconfigure(1, weight=1)
            ttk.Label(parent_f, text=label, font=("Malgun Gothic", 8)).grid(row=row, column=0, sticky='e', padx=2, pady=2)
            ttk.Entry(parent_f, textvariable=var, font=("Arial", 9), width=1, exportselection=False).grid(row=row, column=1, padx=2, pady=2, sticky='ew')
            cmd = (lambda: self._browse_dir(var)) if is_dir else (lambda: self._browse_file(var, types))
            ttk.Button(parent_f, text="...", width=3, command=cmd).grid(row=row, column=2, padx=2, pady=2)

        _add_file_row(file_frame, "로고 폴더:", self.logo_folder_path, 0, is_dir=True)
        _add_file_row(file_frame, "PT 데이터:", self.pt_target_file_path, 1, types=[("Excel Source", "*.xls;*.xlsx;*.xlsm")])
        _add_file_row(file_frame, "PT 양식:", self.pt_template_file_path, 2, types=[("Excel Template", "*.xlsx;*.xlsm")])

        # 2. Configuration Tabs
        pt_config_frame = ttk.LabelFrame(left_pane, text=" 리포트 세부 설정 ", padding=2)
        pt_config_frame.pack(fill='both', expand=True, pady=(0, 10))

        self.pt_tab_notebook = ttk.Notebook(pt_config_frame)
        self.pt_tab_notebook.pack(fill='both', expand=True)

        pt_tab_cover = ttk.Frame(self.pt_tab_notebook, padding=5)
        pt_tab_data = ttk.Frame(self.pt_tab_notebook, padding=5)
        pt_tab_rows = ttk.Frame(self.pt_tab_notebook, padding=5)
        pt_tab_cols = ttk.Frame(self.pt_tab_notebook, padding=5)
        
        # [CRITICAL] Allow children to expand
        for t in [pt_tab_cover, pt_tab_data, pt_tab_rows, pt_tab_cols]: t.columnconfigure(0, weight=1)

        self.pt_tab_notebook.add(pt_tab_cover, text="갑지")
        self.pt_tab_notebook.add(pt_tab_data, text="을지")
        self.pt_tab_notebook.add(pt_tab_rows, text="행 설정")
        self.pt_tab_notebook.add(pt_tab_cols, text="컬럼 설정")

        next_row_pt_cover = self._create_setting_grid(pt_tab_cover, "COVER")
        next_row_pt_data = self._create_setting_grid(pt_tab_data, "DATA")
        self._create_margin_settings(pt_tab_cover, "COVER", next_row_pt_cover)
        self._create_margin_settings(pt_tab_data, "DATA", next_row_pt_data)
        self._create_row_settings(pt_tab_rows, mode="PT")
        
        pt_items = [
            ("순번(No):", "PT_COL_NO", 1, "PT_NAME_NO", "No", "No"),
            ("ISO/Dwg:", "PT_COL_DWG", 2, "PT_NAME_DWG", "Dwg", "Dwg"),
            ("Joint No:", "PT_COL_JOINT", 5, "PT_NAME_JOINT", "Joint", "Joint"),
            ("NPS:", "PT_COL_NPS", 6, "PT_NAME_NPS", "NPS", "NPS"),
            ("두께(Th'k):", "PT_COL_THK", 7, "PT_NAME_THK", "Thk.", "Thk."),
            ("재질(Material):", "PT_COL_MAT", 8, "PT_NAME_MAT", "Material", "Material"),
            ("용접사(Welder):", "PT_COL_WELDER", 9, "PT_NAME_WELDER", "Welder", "Welder"),
            ("검사타입(Type):", "PT_COL_TYPE", 10, "PT_NAME_TYPE", "WType", "WType"),
            ("판정 결과:", "PT_COL_RES", 11, "PT_NAME_RES", "Result", "Result")
        ]
        self._create_column_mapping_ui(pt_tab_cols, "PT", pt_items)
        

        # 3. Action Section
        action_frame = tk.Frame(left_pane, background="#ffffff", highlightthickness=1, highlightbackground="#d1d5db", padx=10, pady=5)
        action_frame.pack(fill='x', pady=(0, 10))

        filter_row = tk.Frame(action_frame, background="#ffffff")
        filter_row.pack(fill='x', pady=2)
        ttk.Label(filter_row, text="📊 특정순번:", background="#ffffff", font=("Malgun Gothic", 8)).pack(side='left')
        ttk.Entry(filter_row, textvariable=self.sequence_filter, width=15, font=("Arial", 9), exportselection=False).pack(side='left', padx=5, fill='x', expand=True)

        btn_row = tk.Frame(left_pane, background="#f9fafb")
        btn_row.pack(fill='x', pady=5)
        ttk.Button(btn_row, text=" ✨ 성적서 생성 ", style="Action.TButton", command=self.run_process).pack(fill='x', pady=(0, 5))
        ttk.Button(btn_row, text=" 📝 데이터 추출 ", command=self.extract_only).pack(fill='x')

        # [RIGHT] Preview Pane
        right_frame = ttk.LabelFrame(self.pt_paned, text=" 🔬 실시간 데이터 관리 (PT Preview) ", padding=10)
        self.pt_paned.add(right_frame, stretch="always")
        self._create_pt_preview_ui(right_frame)
        self._apply_sash_ratio("PT")
        

        # Adaptive Resizing Bindings
        self.pt_pane_ratio = self.config.get('PT_SASH_RATIO', 0.5)
        self.pt_paned.bind("<Configure>", lambda e: [self._on_pt_paned_configure(e), self.root.update_idletasks()])
        self.pt_paned.bind("<ButtonRelease-1>", lambda e: self.root.after(10, self._update_pt_ratio))
        self.root.after(500, lambda: self._on_pt_paned_configure(None))

    def _update_pt_ratio(self):
        try:
            total_w = self.pt_paned.winfo_width()
            if total_w > 100:
                current_sash = self.pt_paned.sash_coord(0)[0]
                self.pt_pane_ratio = current_sash / total_w
                self.config['PT_SASH_RATIO'] = self.pt_pane_ratio
                self.save_settings()
        except: pass

    def _on_pt_paned_configure(self, event):
        try:
            total_w = self.pt_paned.winfo_width()
            if total_w > 100:
                if event and event.widget == self.pt_paned:
                    new_pos = int(total_w * self.pt_pane_ratio)
                    self.pt_paned.sash_place(0, new_pos, 0)
        except: pass

    def _create_pt_preview_ui(self, parent):
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # [NEW] File Info Header
        header_info = tk.Frame(container, background="#ffffff", highlightthickness=1, highlightbackground="#e5e7eb")
        header_info.pack(fill='x', pady=(0, 5))
        tk.Label(header_info, textvariable=self.file_info_vars['PT'], background="#ffffff", 
                 foreground="#4b5563", font=("Malgun Gothic", 8, "bold"), padx=10, pady=2).pack(side='left')

        self.pt_display_cols = ["V", "No", "Date", "ISO/Dwg", "Joint No.", "Material", "Test Item", "Result", "Welder No", "Remarks"]
        saved_widths = self.config.get("PT_COL_WIDTHS", {})
        default_widths = {"V": 40, "No": 50, "Date": 90, "ISO/Dwg": 300, "Joint No.": 120, "Material": 100, "Test Item": 100, "Result": 80, "Welder No": 100}

        tree_frame = tk.Frame(container, background="#f9fafb")
        self.pt_preview_tree = ttk.Treeview(tree_frame, columns=self.pt_display_cols, show='headings', height=10, selectmode='extended')
        for col in self.pt_preview_tree["columns"]:
            name_key = f"PT_NAME_{col.split('(')[0].replace(' ', '').replace('.', '').upper()}"
            if col == "Date": name_key = "PT_NAME_DATE"
            elif col == "ISO/Dwg": name_key = "PT_NAME_DWG"
            elif col == "Joint No.": name_key = "PT_NAME_JOINT"
            elif col == "Material": name_key = "PT_NAME_MAT"
            elif col == "Thk": name_key = "PT_NAME_THK"
            elif col == "Test Item": name_key = "PT_NAME_ITEM"
            elif col == "Result": name_key = "PT_NAME_RES"
            elif col == "Welder No": name_key = "PT_NAME_WELDER"
            else: name_key = None
            
            display_text = self.config.get(name_key, col) if name_key else col
            self.pt_preview_tree.heading(col, text=display_text, anchor='center', command=lambda _c=col: self.sort_by_column(_c, mode="PT"))
            w = saved_widths.get(col, default_widths.get(col, 80))
            self.pt_preview_tree.column(col, width=w, anchor='center', stretch=False)
        
        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.pt_preview_tree.yview)
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.pt_preview_tree.xview)
        self.pt_preview_tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.pt_preview_tree.grid(row=0, column=0, sticky='nsew')
        scroll_y.grid(row=0, column=1, sticky='ns')
        scroll_x.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        self._setup_preview_sidebar(self.pt_preview_tree, container, mode="PT")
        tree_frame.pack(side="left", fill="both", expand=True)

    def _setup_paut_ui(self, parent):
        # [FORCE] Ensure parent (paut_mode_frame) allows expansion
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # --- Dual Pane Layout (PanedWindow) ---
        self.paut_paned = tk.PanedWindow(container, orient='horizontal', background="#d1d5db", 
                            sashwidth=6, sashpad=0, sashrelief='raised', borderwidth=0)
        self.paut_paned.pack(fill='both', expand=True)

        # [LEFT] Settings & Manual Eval Pane
        left_pane = tk.Frame(self.paut_paned, background="#f9fafb", padx=10, pady=10)
        self.paut_paned.add(left_pane, stretch="always")

        # Header
        header_frame = tk.Frame(left_pane, background="#f9fafb")
        header_frame.pack(fill='x', pady=(0, 10))
        tk.Label(header_frame, text="🔬 PAUT (ASME B31.1) 관리", font=("Malgun Gothic", 15, "bold"), 
                 background="#f9fafb", foreground="#1e3a8a").pack(side='left')

        # 1. File Selection Group
        file_frame = ttk.LabelFrame(left_pane, text=" 데이터 및 양식 (Data) ", padding=10)
        file_frame.pack(fill='x', pady=(0, 10))

        def _add_file_row(parent_f, label, var, row, is_dir=False, types=None):
            parent_f.columnconfigure(1, weight=1)
            ttk.Label(parent_f, text=label, font=("Malgun Gothic", 8)).grid(row=row, column=0, sticky='e', padx=2, pady=2)
            ttk.Entry(parent_f, textvariable=var, font=("Arial", 9), width=1, exportselection=False).grid(row=row, column=1, padx=2, pady=2, sticky='ew')
            cmd = (lambda: self._browse_dir(var)) if is_dir else (lambda: self._browse_file(var, types))
            ttk.Button(parent_f, text="...", width=3, command=cmd).grid(row=row, column=2, padx=2, pady=2)

        _add_file_row(file_frame, "PAUT 데이터:", self.paut_target_file_path, 0, types=[("Excel Source", "*.xls;*.xlsx;*.xlsm")])
        _add_file_row(file_frame, "PAUT 양식:", self.paut_template_file_path, 1, types=[("Excel Template", "*.xlsx;*.xlsm")])

        # --- 2. Configuration Notebook (갑, 을, 행 설정, 개별 판정) ---
        config_frame = ttk.LabelFrame(left_pane, text=" 설정 및 판정 (Config & Eval) ", padding=2)
        config_frame.pack(fill='both', expand=True, pady=(0, 10))

        self.paut_tab_notebook = ttk.Notebook(config_frame)
        self.paut_tab_notebook.pack(fill='both', expand=True)

        tab_cover = ttk.Frame(self.paut_tab_notebook, padding=5)
        tab_data = ttk.Frame(self.paut_tab_notebook, padding=5)
        tab_rows = ttk.Frame(self.paut_tab_notebook, padding=5)
        tab_cols = ttk.Frame(self.paut_tab_notebook, padding=5)
        tab_eval = ttk.Frame(self.paut_tab_notebook, padding=5)

        # Allow expansion
        for t in [tab_cover, tab_data, tab_rows, tab_cols, tab_eval]: t.columnconfigure(0, weight=1)

        self.paut_tab_notebook.add(tab_eval, text="개별 판정") # Default/Main tab
        self.paut_tab_notebook.add(tab_cover, text="갑지")
        self.paut_tab_notebook.add(tab_data, text="을지")
        self.paut_tab_notebook.add(tab_rows, text="행 설정")
        self.paut_tab_notebook.add(tab_cols, text="컬럼 설정")

        # [NEW] Populate Settings
        next_row_cover = self._create_setting_grid(tab_cover, "PAUT_COVER")
        next_row_data = self._create_setting_grid(tab_data, "PAUT_DATA")
        self._create_margin_settings(tab_cover, "PAUT_COVER", next_row_cover)
        self._create_margin_settings(tab_data, "PAUT_DATA", next_row_data)
        self._create_row_settings(tab_rows, mode="PAUT")
        
        paut_items = [
            ("순번(No):", "PAUT_COL_NO", 1, "PAUT_COL_NO_NAME", "No", "No"),
            ("Line No.:", "PAUT_COL_LINE", 2, "PAUT_NAME_LINE", "Line No.", "Line No."),
            ("Joint No.:", "PAUT_COL_JOINT", 3, "PAUT_NAME_JOINT", "Joint No.", "Joint No."),
            ("두께(Th'k):", "PAUT_COL_THK", 4, "PAUT_NAME_THK", "Th'k(mm)", "Th'k(mm)"),
            ("결함높이(H):", "PAUT_COL_H", 5, "PAUT_NAME_H", "Height(mm)", "Height(mm)"),
            ("결함길이(L):", "PAUT_COL_L", 6, "PAUT_NAME_L", "Length(mm)", "Length(mm)"),
            ("결함깊이(D):", "PAUT_COL_UP", 7, "PAUT_NAME_UP", "Upper", "Upper"),
            ("하부깊이(Low):", "PAUT_COL_LOW", 0, "PAUT_NAME_LOW", "Lower", "Lower"),
            ("시작위치:", "PAUT_COL_START", 0, "PAUT_NAME_START", "Start", "Start"),
            ("종료위치:", "PAUT_COL_END", 0, "PAUT_NAME_END", "End", "End"),
            ("a/l 비:", "PAUT_COL_AL", 0, "PAUT_NAME_AL", "a/l", "a/l"),
            ("a/t 비:", "PAUT_COL_AT", 0, "PAUT_NAME_AT", "a/t", "a/t"),
            ("결함종류:", "PAUT_COL_NAT", 8, "PAUT_NAME_NAT", "Type of Flaw", "Type of Flaw"),
            ("판정(Eval):", "PAUT_COL_EVAL", 9, "PAUT_NAME_EVAL", "Evaluation", "Evaluation"),
            ("비고(Rem):", "PAUT_COL_REM", 10, "PAUT_NAME_REM", "Remarks", "Remarks")
        ]
        self._create_column_mapping_ui(tab_cols, "PAUT", paut_items)

        # --- 3. Manual Evaluation Content (Inside tab_eval) ---
        manual_container = tk.Frame(tab_eval, background="#f9fafb")
        manual_container.pack(fill='both', expand=True)
        
        # [NEW] Evaluation Mode Selector
        mode_frame = tk.Frame(manual_container, background="#f9fafb")
        mode_frame.pack(fill='x', pady=(0, 5))
        ttk.Label(mode_frame, text="✅ 판정 기준:", font=("Malgun Gothic", 8, "bold")).pack(side='left', padx=2)
        tk.Radiobutton(mode_frame, text="일반(B31.1)", variable=self.paut_eval_mode, value="Standard", background="#f9fafb").pack(side='left', padx=5)
        tk.Radiobutton(mode_frame, text="ECA(파과역학)", variable=self.paut_eval_mode, value="ECA", background="#f9fafb").pack(side='left', padx=5)

        # Sizing Helper Row
        calc_frame = tk.Frame(manual_container, background="#f1f5f9", padx=10, pady=8, highlightbackground="#cbd5e1", highlightthickness=1)
        calc_frame.pack(fill='x', pady=(0, 10))
        
        # Row 0: H-Calc Settings
        h_set_row = tk.Frame(calc_frame, background="#f1f5f9")
        h_set_row.pack(fill='x')
        ttk.Label(h_set_row, text="📏 H-Setting:", font=("Malgun Gothic", 8, "bold"), background="#f1f5f9").pack(side='left', padx=2)
        tk.Entry(h_set_row, textvariable=self.paut_manual_vars['peak'], width=5).pack(side='left', padx=2)
        ttk.Label(h_set_row, text="% /", background="#f1f5f9").pack(side='left')
        tk.Entry(h_set_row, textvariable=self.paut_manual_vars['db'], width=4).pack(side='left', padx=2)
        ttk.Label(h_set_row, text="dB ->", background="#f1f5f9").pack(side='left', padx=2)
        ttk.Label(h_set_row, textvariable=self.paut_manual_vars['target_fsh'], background="#f1f5f9", foreground="#ef4444", font=("Arial", 8, "bold")).pack(side='left', padx=2)
        
        # Row 1: H-Calc Data
        h_data_row = tk.Frame(calc_frame, background="#f1f5f9")
        h_data_row.pack(fill='x', pady=2)
        ttk.Label(h_data_row, text="📏 H-Data:", font=("Malgun Gothic", 8, "bold"), background="#f1f5f9").pack(side='left', padx=2)
        tk.Entry(h_data_row, textvariable=self.paut_manual_vars['z1'], width=6).pack(side='left', padx=2)
        ttk.Label(h_data_row, text="~", background="#f1f5f9").pack(side='left')
        tk.Entry(h_data_row, textvariable=self.paut_manual_vars['z2'], width=6).pack(side='left', padx=2)
        ttk.Button(h_data_row, text="입력", width=6, command=self._calculate_paut_h).pack(side='right', padx=5)

        # Row 2: L-Calc Data
        l_data_row = tk.Frame(calc_frame, background="#f1f5f9")
        l_data_row.pack(fill='x')
        ttk.Label(l_data_row, text="📏 L-Data:", font=("Malgun Gothic", 8, "bold"), background="#f1f5f9").pack(side='left', padx=2)
        tk.Entry(l_data_row, textvariable=self.paut_manual_vars['L1'], width=6).pack(side='left', padx=2)
        ttk.Label(l_data_row, text="~", background="#f1f5f9").pack(side='left')
        tk.Entry(l_data_row, textvariable=self.paut_manual_vars['L2'], width=6).pack(side='left', padx=2)
        ttk.Button(l_data_row, text="입력", width=6, command=self._calculate_paut_l).pack(side='right', padx=5)

        # Input Grid
        input_grid = tk.Frame(manual_container, background="#f9fafb")
        input_grid.pack(fill='x', pady=5)
        
        m_inputs = [("두께(T):", "t"), ("높이(H):", "h"), ("길이(L):", "l"), ("깊이(d):", "d"), 
                    ("시작(S):", "L1"), ("종료(E):", "L2"), ("상단(Up):", "D_Upper"), ("하단(Low):", "D_Lower")]
        for i, (lbl, key) in enumerate(m_inputs):
            grid_row = i // 2
            grid_col = (i % 2) * 2
            ttk.Label(input_grid, text=lbl, font=("Malgun Gothic", 8)).grid(row=grid_row, column=grid_col, sticky='e', padx=2, pady=1)
            ent = tk.Entry(input_grid, textvariable=self.paut_manual_vars[key], width=10)
            ent.grid(row=grid_row, column=grid_col+1, padx=2, pady=1, sticky='w')
            ent.bind("<Return>", lambda e: self.root.focus_set())

        # Nature & Evaluation Actions
        nature_row = tk.Frame(manual_container, background="#f9fafb")
        nature_row.pack(fill='x', pady=5)
        ttk.Label(nature_row, text="종류:").pack(side='left', padx=2)
        ttk.Combobox(nature_row, textvariable=self.paut_manual_vars['nature'], values=["Crack", "LOF", "IP", "Slag", "Porosity", "Others"], width=9).pack(side='left', padx=2)
        ttk.Label(nature_row, text="위치:").pack(side='left', padx=2)
        ttk.Label(nature_row, textvariable=self.paut_manual_vars['loc'], foreground="#3b82f6", font=("Arial", 9, "bold")).pack(side='left', padx=2)
        
        btn_row = tk.Frame(manual_container, background="#f9fafb")
        btn_row.pack(fill='x', pady=2)
        ttk.Button(btn_row, text="판정", width=8, command=self._run_manual_paut_eval).pack(side='left', expand=True, fill='x', padx=2)
        ttk.Button(btn_row, text="저장", width=8, command=self._save_paut_manual_to_tree).pack(side='left', expand=True, fill='x', padx=2)

        self.paut_res_label = tk.Label(manual_container, text="결과 대기", font=("Malgun Gothic", 10, "bold"), background="#f3f4f6")
        self.paut_res_label.pack(fill='x', pady=(5, 0))

        # Bottom Actions
        paut_btn_row = tk.Frame(left_pane, background="#f9fafb")
        paut_btn_row.pack(fill='x', pady=(5, 2))
        ttk.Button(paut_btn_row, text=" 📝 추출 ", command=self._extract_paut_data).pack(side='left', fill='x', expand=True, padx=(0, 5))
        ttk.Button(paut_btn_row, text=" ✨ 일괄 판정 ", command=self._run_batch_paut_eval).pack(side='left', fill='x', expand=True, padx=(0, 5))
        ttk.Button(paut_btn_row, text=" 📄 성적서 생성 ", command=self._generate_paut_report).pack(side='left', fill='x', expand=True)

        paut_session_row = tk.Frame(left_pane, background="#f9fafb")
        paut_session_row.pack(fill='x', pady=(0, 10))
        ttk.Button(paut_session_row, text=" 💾 세션 저장 ", command=self._export_paut_session).pack(side='left', fill='x', expand=True, padx=(0, 5))
        ttk.Button(paut_session_row, text=" 📂 세션 불러오기 ", command=self._import_paut_session).pack(side='left', fill='x', expand=True)

        # [RIGHT] Batch & Preview Pane
        right_frame = ttk.LabelFrame(self.paut_paned, text=" 🔬 일괄 판정 및 미리보기 (Batch & Preview) ", padding=10)
        self.paut_paned.add(right_frame, stretch="always")
        self._apply_sash_ratio("PAUT")
        
        # [NEW] File Info Header
        header_info = tk.Frame(right_frame, background="#ffffff", highlightthickness=1, highlightbackground="#e5e7eb")
        header_info.pack(fill='x', pady=(0, 5))
        tk.Label(header_info, textvariable=self.file_info_vars['PAUT'], background="#ffffff", 
                 foreground="#4b5563", font=("Malgun Gothic", 8, "bold"), padx=10, pady=2).pack(side='left')

        tree_frame = tk.Frame(right_frame, background="#f9fafb")
        tree_frame.pack(fill='both', expand=True)
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        self.paut_preview_tree = ttk.Treeview(tree_frame, columns=("V", "No", "Line No.", "Joint No.", "Th'k(mm)", "Start", "End", "Length(mm)", "Upper", "Lower", "Height(mm)", "Type of Flaw", "a/l", "a/t", "Evaluation", "Remarks"), show='headings', height=10, selectmode='extended')
        
        saved_widths = self.config.get("PAUT_COL_WIDTHS", {})
        default_widths = {"V": 40, "No": 50, "Line No.": 250, "Joint No.": 120, "Th'k(mm)": 60, "Start": 60, "End": 60, "Length(mm)": 80, "Upper": 60, "Lower": 60, "Height(mm)": 80, "Type of Flaw": 100, "a/l": 60, "a/t": 60, "Evaluation": 80, "Remarks": 150}
        
        for col in self.paut_preview_tree["columns"]:
            # Use dynamic names from config for headings
            name_key = f"PAUT_NAME_{col.split('(')[0].replace(' ', '').replace('.', '').replace('/', '').upper()}"
            default_name = col
            if col == "No": name_key = "PAUT_COL_NO_NAME"
            elif col == "Line No.": name_key = "PAUT_NAME_LINE"
            elif col == "Joint No.": name_key = "PAUT_NAME_JOINT"
            elif col == "Th'k(mm)": name_key = "PAUT_NAME_THK"
            elif col == "Height(mm)": name_key = "PAUT_NAME_H"
            elif col == "Length(mm)": name_key = "PAUT_NAME_L"
            elif col == "Upper": name_key = "PAUT_NAME_UP"
            elif col == "Lower": name_key = "PAUT_NAME_LOW"
            elif col == "Start": name_key = "PAUT_NAME_START"
            elif col == "End": name_key = "PAUT_NAME_END"
            elif col == "a/l": name_key = "PAUT_NAME_AL"
            elif col == "a/t": name_key = "PAUT_NAME_AT"
            elif col == "Type of Flaw": name_key = "PAUT_NAME_NAT"
            elif col == "Evaluation": name_key = "PAUT_NAME_EVAL"
            elif col == "Remarks": name_key = "PAUT_NAME_REM"
            else: name_key = None
            
            display_text = self.config.get(name_key, default_name) if name_key else default_name
            self.paut_preview_tree.heading(col, text=display_text, anchor='center', command=lambda _c=col: self.sort_by_column(_c, mode="PAUT"))
            
            w = saved_widths.get(col, default_widths.get(col, 80))
            self.paut_preview_tree.column(col, width=w, anchor='center', stretch=False)
        
        # Add Sidebar first
        self._setup_preview_sidebar(self.paut_preview_tree, right_frame, mode="PAUT")
        
        # Then pack tree frame to take remaining space
        tree_frame.pack(side='left', fill='both', expand=True)
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        self.paut_preview_tree.grid(row=0, column=0, sticky='nsew')
        
        paut_vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.paut_preview_tree.yview)
        paut_hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.paut_preview_tree.xview)
        self.paut_preview_tree.configure(yscrollcommand=paut_vsb.set, xscrollcommand=paut_hsb.set)
        
        paut_vsb.grid(row=0, column=1, sticky='ns')
        paut_hsb.grid(row=1, column=0, sticky='ew')


        # Adaptive Resizing Bindings
        self.paut_pane_ratio = self.config.get('PAUT_SASH_RATIO', 0.5)
        self.paut_paned.bind("<Configure>", lambda e: [self._on_paut_paned_configure(e), self.root.update_idletasks()])
        self.paut_paned.bind("<ButtonRelease-1>", lambda e: self.root.after(10, self._update_paut_ratio))
        self.root.after(500, lambda: self._on_paut_paned_configure(None))

    def _update_paut_ratio(self):
        try:
            total_w = self.paut_paned.winfo_width()
            if total_w > 100:
                current_sash = self.paut_paned.sash_coord(0)[0]
                self.paut_pane_ratio = current_sash / total_w
                self.config['PAUT_SASH_RATIO'] = self.paut_pane_ratio
                self.save_settings()
        except: pass

    def _on_paut_paned_configure(self, event):
        try:
            total_w = self.paut_paned.winfo_width()
            if total_w > 100:
                if event and event.widget == self.paut_paned:
                    new_pos = int(total_w * self.paut_pane_ratio)
                    self.paut_paned.sash_place(0, new_pos, 0)
        except: pass

    def _update_paut_auto_loc(self):
        try:
            t = float(self.paut_manual_vars['t'].get() or 0)
            d = float(self.paut_manual_vars['d'].get() or 0)
            h = float(self.paut_manual_vars['h'].get() or 0)
            if t > 0:
                is_surface = (d <= 0.1 * t) or (d + h >= 0.9 * t)
                loc = "Surface" if is_surface else "Subsurface"
                self.paut_manual_vars['loc'].set(loc)
        except: pass

    def _update_paut_target_fsh(self):
        """[NEW] Peak 나 dB 변경 시 목표 % FSH 실시간 업데이트"""
        try:
            peak = float(self.paut_manual_vars['peak'].get() or 0)
            db = float(self.paut_manual_vars['db'].get() or 0)
            if peak > 0:
                drop_factor = pow(10, -db / 20)
                target = peak * drop_factor
                self.paut_manual_vars['target_fsh'].set(f"{target:.1f}%")
            else:
                self.paut_manual_vars['target_fsh'].set("0.0%")
        except:
            self.paut_manual_vars['target_fsh'].set("-")

    def _calculate_paut_h(self):
        """[NEW] dB 강하법을 이용한 h 값 계산 및 자동 입력"""
        try:
            peak = float(self.paut_manual_vars['peak'].get() or 80)
            db = float(self.paut_manual_vars['db'].get() or 6)
            z1 = float(self.paut_manual_vars['z1'].get() or 0)
            z2 = float(self.paut_manual_vars['z2'].get() or 0)
            
            # Calculate Target FSH: Peak * 10^(-dB/20)
            drop_factor = pow(10, -db / 20)
            target_fsh = peak * drop_factor
            self.paut_manual_vars['target_fsh'].set(f"{target_fsh:.1f}%")

            h_calc = abs(z1 - z2)
            self.paut_manual_vars['h'].set(f"{h_calc:.2f}")
            
            # [NEW] D: 중간값, UP: 낮은값, LOW: 높은값 자동 기입
            d_mid = (z1 + z2) / 2
            d_min = min(z1, z2)
            d_max = max(z1, z2)
            
            self.paut_manual_vars['d'].set(f"{d_mid:.2f}")
            self.paut_manual_vars['D_Upper'].set(f"{d_min:.2f}")
            self.paut_manual_vars['D_Lower'].set(f"{d_max:.2f}")

            self._update_paut_auto_loc()
            self.log(f"📐 h 산출 완료 ({peak}% peak -> {db}dB 강하 목표 {target_fsh:.1f}%): {h_calc:.2f} mm (D:{d_mid:.2f}, U:{d_min:.2f}, L:{d_max:.2f})")
        except Exception as e:
            messagebox.showerror("계산 오류", f"입력값을 확인해주세요: {e}")

    def _calculate_paut_l(self):
        """[NEW] L1, L2 값을 이용한 결함 길이(L) 자동 계산"""
        try:
            l1 = float(self.paut_manual_vars['L1'].get() or 0)
            l2 = float(self.paut_manual_vars['L2'].get() or 0)
            l_calc = abs(l1 - l2)
            self.paut_manual_vars['l'].set(f"{l_calc:.2f}")
            self.log(f"📐 L 산출 완료 (L1:{l1} ~ L2:{l2}): {l_calc:.2f} mm")
        except Exception as e:
            messagebox.showerror("계산 오류", f"L1, L2 입력값을 확인해주세요: {e}")

    def _save_paut_manual_to_tree(self):
        """[NEW] 수동 판정 구역의 입력값을 선택된 트리뷰 행(들)에 저장"""
        selected_ids = self.paut_preview_tree.selection()
        if not selected_ids:
            messagebox.showwarning("선택 누락", "저장할 행을 먼저 선택해주세요.")
            return

        try:
            t = self.to_float(self.paut_manual_vars['t'].get() or 0)
            h = self.to_float(self.paut_manual_vars['h'].get() or 0)
            l = self.to_float(self.paut_manual_vars['l'].get() or 0)
            d = self.to_float(self.paut_manual_vars['d'].get() or 0)
            l1 = self.paut_manual_vars['L1'].get()
            l2 = self.paut_manual_vars['L2'].get()
            up = self.paut_manual_vars['D_Upper'].get()
            low = self.paut_manual_vars['D_Lower'].get()
            nat = self.paut_manual_vars['nature'].get()
            rem = "" # Remarks can be added to UI if needed, for now keep existing or clear

            # Run evaluation logic to get the result
            eval_mode = self.paut_eval_mode.get()
            res, loc = self.evaluate_paut_flaw(t, h, l, d, nat, mode=eval_mode)
            
            al = f"{h/l:.3f}" if l > 0 else ""
            at = f"{h/t:.3f}" if t > 0 else ""

            mode_info = self._get_mode_info("PAUT")
            if not mode_info: return
            _, idx_map, data_list, _ = mode_info

            for item_id in selected_ids:
                view_idx = self.paut_preview_tree.index(item_id)
                if 0 <= view_idx < len(idx_map):
                    actual_idx = idx_map[view_idx]
                    item = data_list[actual_idx]
                    item["Th'k(mm)"] = t
                    item['Height(mm)'] = h
                    item['Length(mm)'] = l
                    item['Start'] = l1
                    item['End'] = l2
                    item['Upper'] = up
                    item['Lower'] = low
                    item['Type of Flaw'] = nat
                    item['Evaluation'] = f"{res} ({loc})"
                    item['a/l'] = al
                    item['a/t'] = at
            
            self.populate_preview(data_list, mode="PAUT", switch_tab=False)
            self.log(f"💾 선택된 {len(selected_ids)}개 행에 수동 판정 결과 저장 완료.")
            
            # [NEW] 저장 후 입력 필드 초기화
            self._reset_paut_manual_inputs()
            
        except Exception as e:
            messagebox.showerror("저장 오류", f"입력값을 확인해주세요: {e}")

    def _run_manual_paut_eval(self):
        try:
            t = float(self.paut_manual_vars['t'].get() or 0)
            h = float(self.paut_manual_vars['h'].get() or 0)
            l = float(self.paut_manual_vars['l'].get() or 0)
            d = float(self.paut_manual_vars['d'].get() or 0)
            nat = self.paut_manual_vars['nature'].get()
            eval_mode = self.paut_eval_mode.get()
            
            res, loc = self.evaluate_paut_flaw(t, h, l, d, nat, mode=eval_mode)
            self.paut_manual_vars['loc'].set(loc)
            if res == "Accept":
                self.paut_res_label.config(text=f"{res} (위치: {loc})", fg="white", bg="#27ae60")
            else:
                self.paut_res_label.config(text=f"{res} (위치: {loc})", fg="white", bg="#e74c3c")
        except Exception as e:
            messagebox.showerror("입력 오류", f"입력값을 확인해주세요: {e}")


    def _extract_paut_data(self):
        file_path = self.paut_target_file_path.get()
        if not file_path:
            messagebox.showwarning("파일 미선택", "PAUT 데이터 엑셀 파일을 먼저 선택해주세요.")
            return
            
        try:
            self.log(f"📂 PAUT 데이터 불러오는 중: {os.path.basename(file_path)}")
            self.progress['value'] = 20
            
            df = pd.read_excel(file_path)
            cols = df.columns.tolist()
            
            mapping = {
                "t": next((c for c in cols if any(x in c.upper() for x in ["THK", "두께", "TH'K", "THICKNESS"])), None),
                "h": next((c for c in cols if any(x in c.upper() for x in ["HEIGHT", "높이", "FLAW_HEIGHT"])), None),
                "l": next((c for c in cols if any(x in c.upper() for x in ["LENGTH", "길이", "FLAW_LENGTH", "LENTH"])), None),
                "d": next((c for c in cols if any(x in c.upper() for x in ["DEPTH", "깊이", "FLAW_DEPTH"])), None),
                "nature": next((c for c in cols if any(x in c.upper() for x in ["NAT", "종류", "FLAW_NATURE", "TYPE"])), None),
                "line": next((c for c in cols if any(x in c.upper() for x in ["LINE", "ISO", "DWG", "DRAWING"])), None),
                "joint": next((c for c in cols if any(x in c.upper() for x in ["JOINT", "WELD"])), None),
                "start": next((c for c in cols if any(x in c.upper() for x in ["START", "L1"])), None),
                "end": next((c for c in cols if any(x in c.upper() for x in ["END", "L2"])), None),
                "upper": next((c for c in cols if any(x in c.upper() for x in ["UPPER", "UP", "DEPTH"])), None),
                "lower": next((c for c in cols if any(x in c.upper() for x in ["LOWER", "LOW"])), None),
                "remarks": next((c for c in cols if any(x in c.upper() for x in ["REMARK", "REMAKER", "비고"])), None)
            }
            
            if not all([mapping["t"], mapping["h"], mapping["l"], mapping["d"]]):
                self.log("⚠️ 필수 컬럼을 모두 찾을 수 없어 자동 매핑에 실패했습니다.")
                
            self.paut_extracted_data = []
            for _, row in df.iterrows():
                item = {
                    'selected': True,
                    'Line No.': str(row.get(mapping["line"], "")) if mapping["line"] else "",
                    'Joint No.': str(row.get(mapping["joint"], "")) if mapping["joint"] else "",
                    'Th\'k(mm)': row.get(mapping["t"], 0),
                    'Start': row.get(mapping["start"], ""),
                    'End': row.get(mapping["end"], ""),
                    'Length(mm)': row.get(mapping["l"], 0),
                    'Upper': row.get(mapping["upper"], ""),
                    'Lower': row.get(mapping["lower"], ""),
                    'Height(mm)': row.get(mapping["h"], 0),
                    'Type of Flaw': str(row.get(mapping["nature"], "Slag")) if mapping["nature"] else "Slag",
                    'a/l': "", 'a/t': "",
                    'Evaluation': "",
                    'Remarks': str(row.get(mapping["remarks"], "")) if mapping["remarks"] else ""
                }
                # [NEW] 원본 엑셀의 모든 컬럼 데이터를 유지하여 '컬럼 관리'에서 활용 가능하게 함
                item_full = row.to_dict()
                item_full.update(item)
                # [NEW] 수동 추가된 컬럼(Remark 등)이 추출 시 증발하지 않도록 초기화
                for k in self.paut_column_keys:
                    if k not in item_full and k != "selected":
                        item_full[k] = ""
                self.paut_extracted_data.append(item_full)
            
            self.progress['value'] = 100
            self.sort_by_column("ISO/DWG", mode="PAUT") # Auto sort for PAUT
            self.log(f"✅ PAUT 데이터 추출 완료: {len(self.paut_extracted_data)} 건")
            
        except Exception as e:
            self.log(f"❌ PAUT 데이터 추출 오류: {e}")

    def _run_batch_paut_eval(self):
        if not self.paut_extracted_data:
            messagebox.showwarning("데이터 없음", "먼저 데이터를 추출해주세요.")
            return
            
        self.log("🚀 PAUT 일괄 판정 시작...")
        count_ok, count_ng = 0, 0
        
        for item in self.paut_extracted_data:
            t = self.to_float(item.get('Th\'k(mm)', 0))
            h = self.to_float(item.get('Height(mm)', 0))
            l = self.to_float(item.get('Length(mm)', 0))
            d = self.to_float(item.get('Upper', 0))
            nat = item.get('Type of Flaw', 'Slag')
            eval_mode = self.paut_eval_mode.get()
            
            res, loc = self.evaluate_paut_flaw(t, h, l, d, nat, mode=eval_mode)
            item['Evaluation'] = f"{res} ({loc})"
            # a/l, a/t calculation
            if l > 0: item['a/l'] = f"{h/l:.3f}"
            if t > 0: item['a/t'] = f"{h/t:.3f}"
            
            if res == "Accept": count_ok += 1
            else: count_ng += 1
            
        self.populate_preview(self.paut_extracted_data, mode="PAUT")
        self.log(f"✅ 판정 완료: 총 {len(self.paut_extracted_data)} 건 (합격: {count_ok}, 불합격: {count_ng})")

    def _generate_paut_report(self):
        template_path = self.paut_template_file_path.get()
        if not template_path or not os.path.exists(template_path):
            messagebox.showwarning("파일 미선택", "PAUT 성적서 양식 파일을 선택해주세요.")
            return

        final_list = [d for d in self.paut_extracted_data if d.get('selected', True)]
        if not final_list:
            messagebox.showwarning("항목 미선택", "선택된 데이터가 없습니다.")
            return

        self.log(f"🚀 PAUT 성적서 생성 시작 (총 {len(final_list)} 건)...")
        self.progress['value'] = 0
        
        try:
            wb = openpyxl.load_workbook(template_path, keep_vba=True)
            ws = wb.worksheets[0]
            ws.title = f"PAUT_Report_001"
            
            self.add_logos_to_sheet(ws, is_cover=False)
            self.force_print_settings(ws, context="DATA")
            
            start_row = int(self.config.get('PAUT_START_ROW', 11))
            end_row = int(self.config.get('PAUT_DATA_END_ROW', 40))
            
            curr_row = start_row
            data_font = Font(size=9)
            
            for i, item in enumerate(final_list):
                if curr_row > end_row:
                    ws = self.prepare_next_paut_sheet(wb, ws.title, i//(end_row-start_row+1))
                    curr_row = start_row
                
                mapping = {
                    self.config.get('PAUT_COL_NO', '1'): item.get('No', i+1), 
                    self.config.get('PAUT_COL_LINE', '2'): item.get('Line No.', ''), 
                    self.config.get('PAUT_COL_JOINT', '3'): item.get('Joint No.', ''),
                    self.config.get('PAUT_COL_THK', '4'): item.get('Th\'k(mm)', ''), 
                    self.config.get('PAUT_COL_H', '5'): item.get('Height(mm)', ''), 
                    self.config.get('PAUT_COL_L', '6'): item.get('Length(mm)', ''),
                    self.config.get('PAUT_COL_UP', '7'): item.get('Upper', ''), 
                    self.config.get('PAUT_COL_LOW', '0'): item.get('Lower', ''),
                    self.config.get('PAUT_COL_START', '0'): item.get('Start', ''),
                    self.config.get('PAUT_COL_END', '0'): item.get('End', ''),
                    self.config.get('PAUT_COL_AL', '0'): item.get('a/l', ''),
                    self.config.get('PAUT_COL_AT', '0'): item.get('a/t', ''),
                    self.config.get('PAUT_COL_NAT', '8'): item.get('Type of Flaw', ''), 
                    self.config.get('PAUT_COL_EVAL', '9'): item.get('Evaluation', ''),
                    self.config.get('PAUT_COL_REM', '10'): item.get('Remarks', '')
                }
                
                for col_key, val in mapping.items():
                    try:
                        col_idx = self.col_to_num(col_key)
                        if col_idx < 1: continue # Skip if 0 or invalid
                        
                        cell = ws.cell(row=curr_row, column=col_idx, value=val)
                        cell.font = data_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    except Exception:
                        continue
                
                curr_row += 1
                self.progress['value'] = (i+1) / len(final_list) * 100
            
            out_name = f"PAUT_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            out_path = os.path.join(BASE_DIR, "output", out_name) if os.path.exists(os.path.join(BASE_DIR, "output")) else os.path.join(BASE_DIR, out_name)
            
            if not os.path.exists(os.path.dirname(out_path)): os.makedirs(os.path.dirname(out_path))
            wb.save(out_path)
            self.log(f"✅ PAUT 성적서 생성 완료: {out_name}")
            os.startfile(os.path.dirname(out_path))
            
        except Exception as e:
            self.log(f"❌ PAUT 성적서 생성 오류: {e}")

    def _export_paut_session(self):
        """[NEW] 현재 PAUT 미리보기 데이터를 CSV로 내보내기"""
        if not self.paut_extracted_data:
            messagebox.showwarning("데이터 없음", "내보낼 데이터가 없습니다.")
            return
        
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="PAUT 세션 저장",
            initialfile=f"PAUT_Session_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        if not path: return

        try:
            df = pd.DataFrame(self.paut_extracted_data)
            # UTF-8-SIG for Excel compatibility
            df.to_csv(path, index=False, encoding='utf-8-sig')
            self.log(f"💾 PAUT 세션 저장 완료: {os.path.basename(path)}")
            messagebox.showinfo("저장 완료", f"현재 세션이 저장되었습니다.\n{os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("저장 오류", f"세션 저장 중 오류 발생: {e}")

    def _import_paut_session(self):
        """[NEW] 저장된 PAUT 세션 CSV 불러오기"""
        path = filedialog.askopenfilename(
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="PAUT 세션 불러오기"
        )
        if not path: return

        try:
            df = pd.read_csv(path)
            # Ensure mandatory columns or basic structure
            new_data = df.to_dict('records')
            
            # Simple normalization (handle NaNs from CSV)
            normalized_data = []
            for d in new_data:
                clean_dict = {}
                for k, v in d.items():
                    if pd.isna(v): clean_dict[k] = ""
                    else: clean_dict[k] = v
                normalized_data.append(clean_dict)
            
            self.paut_extracted_data = normalized_data
            self.file_info_vars['PAUT'].set(f"📂 로드됨: {os.path.basename(path)} (총 {len(normalized_data)}건)")
            self.populate_preview(self.paut_extracted_data, mode="PAUT")
            self.log(f"📂 PAUT 세션 불러오기 완료: {os.path.basename(path)} ({len(normalized_data)}건)")
        except Exception as e:
            messagebox.showerror("불러오기 오류", f"세션 불러오기 중 오류 발생: {e}")

    def _reset_paut_manual_inputs(self):
        """[NEW] 개별 판정 입력 필드 초기화 (두께 T 등 공통 설정값 제외)"""
        for k in ['h', 'l', 'd', 'z1', 'z2', 'L1', 'L2', 'D_Upper', 'D_Lower']:
            if k in self.paut_manual_vars:
                self.paut_manual_vars[k].set("")
        self.paut_manual_vars['nature'].set("Slag")
        self.paut_manual_vars['loc'].set("-")
        self.paut_res_label.config(text="결과 대기", fg="black", bg="#f3f4f6")

    def prepare_next_paut_sheet(self, wb, prev_title, page_num):
        source_sheet = wb[prev_title]
        new_sheet = wb.copy_worksheet(source_sheet)
        new_sheet.title = f"PAUT_Report_{page_num+1:03d}"
        
        start_row = int(self.config.get('PAUT_START_ROW', 11))
        end_row = int(self.config.get('PAUT_DATA_END_ROW', 40))
        for r in range(start_row, end_row + 1):
            for c in range(1, 11): new_sheet.cell(row=r, column=c).value = None
        
        self.add_logos_to_sheet(new_sheet, is_cover=False)
        return new_sheet

    def col_to_num(self, col_str):
        """[NEW] 엑셀 열 이름(A, B, AA...)을 숫자 인덱스(1, 2, 27...)로 변환"""
        if not col_str: return 0
        s = str(col_str).strip().upper()
        if s.isdigit(): return int(s)
        
        res = 0
        for i, char in enumerate(reversed(s)):
            if ord('A') <= ord(char) <= ord('Z'):
                res += (ord(char) - ord('A') + 1) * (26 ** i)
            else:
                return 0 # Invalid
        return res

    def _create_tree_context_menu(self):
        """미리보기 트리뷰용 공통 컨텍스트 메뉴 초기화"""
        self.ctx_menu = tk.Menu(self.root, tearoff=0)
        # 기본 항목 기입 (show_context_menu에서 command 동적 바인딩)
        self.ctx_menu.add_command(label="선택 항목 체크/해제 토글 (Toggle Check)")
        self.ctx_menu.add_command(label="현재 셀 내용만 복사 (Copy Cell)")
        self.ctx_menu.add_command(label="선택 행 전체 복사 (Copy Row)")
        self.ctx_menu.add_command(label="이 컬럼 전체 복사 (Copy Column)")
        self.ctx_menu.add_command(label="셀 내용 붙여넣기 (Paste)")
        self.ctx_menu.add_separator()
        self.ctx_menu.add_command(label="위로 이동 (Move Up)")
        self.ctx_menu.add_command(label="아래로 이동 (Move Down)")
        self.ctx_menu.add_separator()
        self.ctx_menu.add_command(label="행 추가 (Add Row)")
        self.ctx_menu.add_command(label="선택 삭제 (Delete Selected)")
        self.ctx_menu.add_separator()
        self.ctx_menu.add_command(label="선택 항목 ISO 병합 (Merge ISO)")
        self.ctx_menu.add_command(label="선택 항목 Joint 병합 (Merge Joint)")
        self.ctx_menu.add_command(label="선택 항목 ISO 병합 해제 (Ungroup ISO)")
        self.ctx_menu.add_command(label="선택 항목 Joint 병합 해제 (Ungroup Joint)")
        self.ctx_menu.add_separator()
        self.ctx_menu.add_command(label="선택 항목 일괄 변경 (Bulk Update)")
        # PAUT 전용 (기본은 숨김)
        self.ctx_menu.add_separator()
        self.ctx_menu.add_command(label="컬럼 설정 (Column Manager)")

    def _create_entry_context_menu(self):
        self.entry_popup = tk.Menu(self.root, tearoff=0)
        self.entry_popup.add_command(label="복사 (Copy)", command=self._copy_entry)
        self.entry_popup.add_command(label="붙여넣기 (Paste)", command=self._paste_entry)
        self.entry_popup.add_separator()
        self.entry_popup.add_command(label="잘라내기 (Cut)", command=self._cut_entry)
        
        self.root.bind_class("Entry", "<Button-3>", self._show_entry_popup)
        self.root.bind_class("TEntry", "<Button-3>", self._show_entry_popup)

    def _show_entry_popup(self, event):
        self.current_entry = event.widget
        self._popup_selection_range = None
        if self.current_entry.select_present():
            self._popup_selection_range = (self.current_entry.index("sel.first"), self.current_entry.index("sel.last"))
        self.entry_popup.tk_popup(event.x_root, event.y_root)

    def _copy_entry(self):
        try:
            if getattr(self, 'current_entry', None):
                if self.current_entry.select_present():
                    first, last = self.current_entry.index("sel.first"), self.current_entry.index("sel.last")
                elif getattr(self, '_popup_selection_range', None):
                    first, last = self._popup_selection_range
                else:
                    return
                text = self.current_entry.get()[first:last]
                self.root.clipboard_clear()
                self.root.clipboard_append(text)
                self.root.update()
        except Exception: pass

    def _paste_entry(self):
        try:
            text = self.root.clipboard_get()
            if getattr(self, 'current_entry', None):
                if getattr(self, '_popup_selection_range', None):
                    first, last = self._popup_selection_range
                    self.current_entry.delete(first, last)
                    self.current_entry.insert(first, text)
                elif self.current_entry.select_present():
                    self.current_entry.delete("sel.first", "sel.last")
                    self.current_entry.insert(tk.INSERT, text)
                else:
                    self.current_entry.insert(tk.INSERT, text)
        except Exception: pass

    def _cut_entry(self):
        self._copy_entry()
        try:
            if getattr(self, 'current_entry', None):
                if getattr(self, '_popup_selection_range', None):
                    first, last = self._popup_selection_range
                    self.current_entry.delete(first, last)
                elif self.current_entry.select_present():
                    self.current_entry.delete("sel.first", "sel.last")
        except Exception: pass

    def _create_setting_grid(self, parent, context):
        # [REFINED] More organized layout for logo settings
        items = [
            ("SITCO 로고", f"SITCO_{context}"), 
            ("서울검사 로고", f"SEOUL_{context}"), 
            ("바닥글 우측 (Footer Right)", f"FOOTER_{context}"), 
            ("바닥글 좌측 (Footer Left)", f"FOOTER_PT_{context}")
        ]
        
        next_row = 0
        for i, (label, key_prefix) in enumerate(items):
            # [HYPER-REACTIVE] Force columnspan=1 to match weighted col 0 of parent
            short_label = label.split(" ")[0].replace("로고", "")
            block = tk.LabelFrame(parent, text=f" {short_label} ", padx=2, pady=1, background="#ffffff", font=("Malgun Gothic", 8, "bold"))
            block.grid(row=i, column=0, sticky='ew', pady=1, padx=2)
            
            frame = tk.Frame(block, background="#ffffff")
            frame.pack(fill='x')
            frame.columnconfigure(1, weight=5, minsize=0) # Path entry weight
            for c in [4, 6, 8, 10]: frame.columnconfigure(c, weight=1, minsize=0) # Coords weight
            
            tk.Label(frame, text="P:", width=2, anchor='e', background="#ffffff", font=("Malgun Gothic", 7)).grid(row=0, column=0)
            v_path = tk.StringVar(value=self.config.get(f"{key_prefix}_PATH", ""))
            ttk.Entry(frame, textvariable=v_path, width=1, exportselection=False).grid(row=0, column=1, sticky='ew', padx=1)
            ttk.Button(frame, text="..", width=2, command=lambda v=v_path: self._browse_file(v, [("Images", "*.png;*.jpg;*.jpeg")])).grid(row=0, column=2, padx=1)
            self.setting_vars[f"{key_prefix}_PATH"] = v_path
            
            for idx, (coord, key_suffix) in enumerate([("X", "X"), ("Y", "Y"), ("W", "W"), ("H", "H")]):
                tk.Label(frame, text=f"{coord}:", width=1, anchor='e', background="#ffffff", font=("Malgun Gothic", 7)).grid(row=0, column=3+idx*2)
                v = tk.StringVar(value=str(self.config.get(f"{key_prefix}_{key_suffix}", "0.0")))
                ttk.Entry(frame, textvariable=v, width=1, exportselection=False).grid(row=0, column=4+idx*2, sticky='ew', padx=1)
                self.setting_vars[f"{key_prefix}_{key_suffix}"] = v
            next_row = i + 1
        return next_row
            
    def _create_margin_settings(self, parent, context, start_row):
        # [HYPER-REACTIVE] Elastic Margin Grid (columnspan=1 for weighted parent)
        frame = ttk.LabelFrame(parent, text=" 인쇄 및 여백 (Margins) ", padding=2)
        frame.grid(row=start_row, column=0, sticky='ew', pady=(5, 0)) 
        
        # Force 0 minsize for elasticity
        for col in range(10): frame.columnconfigure(col, minsize=0)
        # [FIX] Distribute expansion only to entries
        for col in [0, 2, 4, 6, 8]: frame.columnconfigure(col, weight=0) # Labels
        for col in [1, 3, 5, 7, 9]: frame.columnconfigure(col, weight=1) # Entries

        m_items = [("T", "TOP"), ("B", "BOTTOM"), ("L", "LEFT"), ("R", "RIGHT")]
        for i, (sh_lbl, key) in enumerate(m_items):
            ttk.Label(frame, text=f"{sh_lbl}:", font=("Arial", 7)).grid(row=0, column=i*2, sticky='e', padx=1)
            v = tk.StringVar(value=str(self.config.get(f"MARGIN_{context}_{key}", "0.2")))
            ent = ttk.Entry(frame, textvariable=v, width=1)
            ent.grid(row=0, column=i*2+1, sticky='ew', padx=1)
            ent.bind("<Return>", lambda e: self.root.focus_set())
            self.setting_vars[f"MARGIN_{context}_{key}"] = v
            
        ttk.Label(frame, text="S:", font=("Arial", 7)).grid(row=0, column=8, sticky='e', padx=1)
        v_s = tk.StringVar(value=str(self.config.get(f"PRINT_SCALE_{context}", "95")))
        ent_s = ttk.Entry(frame, textvariable=v_s, width=1)
        ent_s.grid(row=0, column=9, sticky='ew', padx=1)
        ent_s.bind("<Return>", lambda e: self.root.focus_set())
        self.setting_vars[f"PRINT_SCALE_{context}"] = v_s

        # Adjusters Row (Proportional weights - only entries grow)
        sub = ttk.Frame(frame)
        sub.grid(row=1, column=0, columnspan=10, sticky='ew', pady=2)
        for c in [1, 3, 5, 7]: sub.columnconfigure(c, weight=1, minsize=0)
        for c in [0, 2, 4, 6]: sub.columnconfigure(c, weight=0, minsize=0)
        
        labels = [("Rows:", f"CUSTOM_ROWS_{context}"), ("H:", f"CUSTOM_ROW_HEIGHT_{context}"), 
                  ("Cols:", f"CUSTOM_COLS_{context}"), ("W:", f"CUSTOM_COL_WIDTH_{context}")]
        
        for i, (lbl, key) in enumerate(labels):
            ttk.Label(sub, text=lbl, font=("Arial", 7)).grid(row=0, column=i*2, sticky='e')
            defaultv = "16.5" if "HEIGHT" in key else ("10.0" if "WIDTH" in key else "")
            v = tk.StringVar(value=str(self.config.get(key, defaultv)))
            ent = ttk.Entry(sub, textvariable=v, width=1)
            ent.grid(row=0, column=i*2+1, sticky='ew', padx=1)
            ent.bind("<Return>", lambda e: self.root.focus_set())
            self.setting_vars[key] = v

    def _create_column_mapping_ui(self, parent, mode, mapping_items):
        """[NEW] 전 모듈 공용 엑셀 컬럼 매핑 및 이름 설정 UI"""
        container = ttk.LabelFrame(parent, text=f" {mode} 성적서 및 미리보기 컬럼 설정 ", padding=10)
        container.pack(fill='both', expand=True, pady=5)
        
        # Grid settings
        # [Label] [Excel Col #] [Display Name]
        container.columnconfigure(2, weight=1)
        container.columnconfigure(5, weight=1)
        
        # Header labels
        ttk.Label(container, text="항목", font=("Malgun Gothic", 8, "bold")).grid(row=0, column=0, sticky='w')
        ttk.Label(container, text="엑셀열 | 표시이름", font=("Malgun Gothic", 8, "bold")).grid(row=0, column=1, columnspan=2, sticky='w')
        ttk.Label(container, text="항목", font=("Malgun Gothic", 8, "bold")).grid(row=0, column=3, sticky='w')
        ttk.Label(container, text="엑셀열 | 표시이름", font=("Malgun Gothic", 8, "bold")).grid(row=0, column=4, columnspan=2, sticky='w')
        
        last_row = 0
        for i, (label, key_idx, def_idx, key_name, def_name, internal_id) in enumerate(mapping_items):
            row = (i // 2) + 1
            last_row = max(last_row, row)
            col_offset = (i % 2) * 3
            
            # 1. Label
            ttk.Label(container, text=label, font=("Malgun Gothic", 8)).grid(row=row, column=col_offset, sticky='e', padx=2, pady=2)
            
            # 2. Excel Col #
            v_idx = tk.StringVar(value=str(self.config.get(key_idx, def_idx)))
            ent_idx = ttk.Entry(container, textvariable=v_idx, width=4)
            ent_idx.grid(row=row, column=col_offset+1, sticky='w', padx=2, pady=2)
            ent_idx.bind("<Return>", lambda e: self.root.focus_set())
            self.setting_vars[key_idx] = v_idx
            
            # 3. Display Name
            v_name = tk.StringVar(value=str(self.config.get(key_name, def_name)))
            ent_name = ttk.Entry(container, textvariable=v_name, width=12)
            ent_name.grid(row=row, column=col_offset+2, sticky='ew', padx=2, pady=2)
            ent_name.bind("<Return>", lambda e: self.root.focus_set())
            self.setting_vars[key_name] = v_name
            
            # Trace to update preview tree heading
            v_name.trace_add("write", lambda *args, tid=internal_id, m=mode, var=v_name: self._update_mode_heading(m, tid, var))

        ttk.Label(container, text="* 엑셀열: 알파벳(A, B...) 또는 숫자(1, 2...) 입력 (0=제외)", foreground="gray", font=("Malgun Gothic", 8)).grid(row=last_row+1, column=0, columnspan=6, pady=10)

    def _update_mode_heading(self, mode, internal_id, var):
        """[NEW] 특정 모드의 미리보기 헤더 이름을 실시간으로 업데이트"""
        try:
            new_text = var.get()
            tree = None
            if mode == "PMI": tree = self.preview_tree
            elif mode == "RT": tree = self.rt_preview_tree
            elif mode == "PT": tree = self.pt_preview_tree
            elif mode == "PAUT": tree = self.paut_preview_tree
            
            if tree:
                tree.heading(internal_id, text=new_text)
        except Exception:
            pass

    def _create_row_settings(self, parent, mode=None):
        if mode is None or mode == "PMI":
            # PMI Rows
            pmi_frame = ttk.LabelFrame(parent, text=" PMI 행 설정 ", padding=10)
            pmi_frame.pack(fill='x', pady=5)
            pmi_rows = [
                ("을지 데이터 시작", "START_ROW", "PMI 데이터 시작"), 
                ("을지 데이터 종료", "DATA_END_ROW", "PMI 데이터 종료"), 
                ("을지 인쇄 영역 끝", "PRINT_END_ROW", "PMI 인쇄 끝"),
                ("갑지 데이터 시작", "GAPJI_START_ROW", "PMI 갑지 데이터 시작/테두리"),
                ("갑지 데이터 종료", "GAPJI_DATA_END_ROW", "PMI 갑지 데이터 종료/테두리"),
                ("갑지 인쇄 영역 끝", "GAPJI_PRINT_END_ROW", "PMI 갑지 인쇄 끝")
            ]
            for i, (label, key, tip) in enumerate(pmi_rows):
                ttk.Label(pmi_frame, text=label + ":").grid(row=i, column=0, sticky='e', padx=5, pady=2)
                # Default values for Gapji settings
                def_val = "0"
                if key == "GAPJI_PRINT_END_ROW": def_val = "51"
                elif key == "GAPJI_START_ROW": def_val = "23"
                elif key == "GAPJI_DATA_END_ROW": def_val = "38"
                else: def_val = str(self.config.get(key, 0))
                
                v = tk.StringVar(value=str(self.config.get(key, def_val)))
                ent = ttk.Entry(pmi_frame, textvariable=v, width=10)
                ent.grid(row=i, column=1, sticky='w', padx=5)
                ent.bind("<Return>", lambda e: self.root.focus_set())
                self.setting_vars[key] = v
                ttk.Label(pmi_frame, text=f"({tip})", foreground="gray", font=("Malgun Gothic", 9)).grid(row=i, column=2, sticky='w')

        if mode is None or mode == "PAUT":
            # PAUT Rows
            ut_row_frame = ttk.LabelFrame(parent, text=" PAUT 행 설정 ", padding=10)
            ut_row_frame.pack(fill='x', pady=5)
            ut_rows = [
                ("을지 데이터 시작", "PAUT_START_ROW", "PAUT 데이터 시작"), 
                ("을지 데이터 종료", "PAUT_DATA_END_ROW", "PAUT 데이터 종료"), 
                ("을지 인쇄 영역 끝", "PAUT_PRINT_END_ROW", "PAUT 인쇄 끝"),
                ("갑지 데이터 시작", "PAUT_GAPJI_START_ROW", "PAUT 갑지 데이터 시작"),
                ("갑지 데이터 종료", "PAUT_GAPJI_DATA_END_ROW", "PAUT 갑지 데이터 종료"),
                ("갑지 인쇄 영역 끝", "PAUT_GAPJI_PRINT_END_ROW", "PAUT 갑지 끝")
            ]
            for i, (label, key, tip) in enumerate(ut_rows):
                ttk.Label(ut_row_frame, text=label + ":").grid(row=i, column=0, sticky='e', padx=5, pady=2)
                def_val = "51" if "GAPJI_PRINT" in key else "0"
                v = tk.StringVar(value=str(self.config.get(key, def_val)))
                ent = ttk.Entry(ut_row_frame, textvariable=v, width=10)
                ent.grid(row=i, column=1, sticky='w', padx=5)
                ent.bind("<Return>", lambda e: self.root.focus_set())
                self.setting_vars[key] = v
                ttk.Label(ut_row_frame, text=f"({tip})", foreground="gray", font=("Malgun Gothic", 9)).grid(row=i, column=2, sticky='w')

        if mode is None or mode == "RT":
            # RT Rows
            rt_row_frame = ttk.LabelFrame(parent, text=" RT 행 설정 ", padding=10)
            rt_row_frame.pack(fill='x', pady=5)
            rt_rows = [
                ("을지 데이터 시작", "RT_START_ROW", "RT 데이터 시작"), 
                ("을지 데이터 종료", "RT_DATA_END_ROW", "RT 데이터 종료"), 
                ("을지 인쇄 영역 끝", "RT_PRINT_END_ROW", "RT 인쇄 끝"),
                ("갑지 데이터 시작", "RT_GAPJI_START_ROW", "RT 갑지 데이터 시작"),
                ("갑지 데이터 종료", "RT_GAPJI_DATA_END_ROW", "RT 갑지 데이터 종료"),
                ("갑지 인쇄 영역 끝", "RT_GAPJI_PRINT_END_ROW", "RT 갑지 끝")
            ]
            for i, (label, key, tip) in enumerate(rt_rows):
                ttk.Label(rt_row_frame, text=label + ":").grid(row=i, column=0, sticky='e', padx=5, pady=2)
                def_val = "51" if "GAPJI_PRINT" in key else "0"
                v = tk.StringVar(value=str(self.config.get(key, def_val)))
                ent = ttk.Entry(rt_row_frame, textvariable=v, width=10)
                ent.grid(row=i, column=1, sticky='w', padx=5)
                ent.bind("<Return>", lambda e: self.root.focus_set())
                self.setting_vars[key] = v
                ttk.Label(rt_row_frame, text=f"({tip})", foreground="gray", font=("Malgun Gothic", 9)).grid(row=i, column=2, sticky='w')

        if mode is None or mode == "PT":
            # PT Rows
            pt_row_frame = ttk.LabelFrame(parent, text=" PT 행 설정 ", padding=10)
            pt_row_frame.pack(fill='x', pady=5)
            pt_rows = [
                ("을지 데이터 시작", "PT_START_ROW", "PT 데이터 시작"), 
                ("을지 데이터 종료", "PT_DATA_END_ROW", "PT 데이터 종료"), 
                ("을지 인쇄 영역 끝", "PT_PRINT_END_ROW", "PT 인쇄 끝"),
                ("갑지 데이터 시작", "PT_GAPJI_START_ROW", "PT 갑지 데이터 시작"),
                ("갑지 데이터 종료", "PT_GAPJI_DATA_END_ROW", "PT 갑지 데이터 종료"),
                ("갑지 인쇄 영역 끝", "PT_GAPJI_PRINT_END_ROW", "PT 갑지 끝")
            ]
            for i, (label, key, tip) in enumerate(pt_rows):
                ttk.Label(pt_row_frame, text=label + ":").grid(row=i, column=0, sticky='e', padx=5, pady=2)
                # Default values for PT
                if key == "PT_START_ROW": def_val = "18"
                elif key == "PT_DATA_END_ROW": def_val = "32"
                elif key == "PT_PRINT_END_ROW": def_val = "35"
                elif key == "PT_GAPJI_PRINT_END_ROW": def_val = "35"
                elif "GAPJI_START" in key: def_val = "0"
                elif "GAPJI_DATA_END" in key: def_val = "35"
                else: def_val = "0"
                
                v = tk.StringVar(value=str(self.config.get(key, def_val)))
                ent = ttk.Entry(pt_row_frame, textvariable=v, width=10)
                ent.grid(row=i, column=1, sticky='w', padx=5)
                ent.bind("<Return>", lambda e: self.root.focus_set())
                self.setting_vars[key] = v
                ttk.Label(pt_row_frame, text=f"({tip})", foreground="gray", font=("Malgun Gothic", 9)).grid(row=i, column=2, sticky='w')

    def _create_preview_ui(self, parent):
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # [NEW] File Info Header (Contextual Awareness)
        header_info = tk.Frame(container, background="#ffffff", highlightthickness=1, highlightbackground="#e5e7eb")
        header_info.pack(fill='x', pady=(0, 5))
        tk.Label(header_info, textvariable=self.file_info_vars['PMI'], background="#ffffff", 
                 foreground="#4b5563", font=("Malgun Gothic", 8, "bold"), padx=10, pady=2).pack(side='left')

        # [REFINED] Compact Header inside Frame
        header_frame = tk.Frame(container, background="#f9fafb")
        header_frame.pack(fill='x', pady=(0, 5))
        
        tk.Checkbutton(header_frame, text="⚠️ 함량 미달만 보기(PMI)", variable=self.pmi_show_deficiency_only, 
                       background="#f9fafb", font=("Malgun Gothic", 9),
                       command=lambda: self.populate_preview(self.extracted_data, switch_tab=False)).pack(side='left')

        tree_frame = tk.Frame(container, background="#f9fafb")
        self.preview_tree = ttk.Treeview(tree_frame, columns=("ST", "V", "No", "Date", "Drawing No.", "Joint No.", "Location", "Ni", "Cr", "Mo", "Result"), show='headings', height=10, selectmode='extended')
        # [NEW] Highlight tags
        self.preview_tree.tag_configure("deficient", background="#fee2e2", foreground="#991b1b") # Light red
        self.preview_tree.tag_configure("group_even", background="#ffffff")
        self.preview_tree.tag_configure("group_odd", background="#f3f4f6")
        
        saved_widths = self.config.get("PMI_COL_WIDTHS", {})
        default_widths = {"ST": 40, "V": 40, "No": 50, "Date": 90, "Drawing No.": 400, "Joint No.": 200, "Location": 300, "Ni": 60, "Cr": 60, "Mo": 60, "Result": 150}
        
        for col in self.preview_tree["columns"]:
            # Use dynamic names from config for headings
            name_key = f"PMI_NAME_{col.split('(')[0].replace(' ', '').replace('.', '').replace('-', '').upper()}"
            if col == "No": name_key = "PMI_NAME_NO"
            elif col == "Date": name_key = "PMI_NAME_DATE"
            elif col == "Drawing No.": name_key = "PMI_NAME_DWG"
            elif col == "Joint No.": name_key = "PMI_NAME_JOINT"
            elif col == "Location": name_key = "PMI_NAME_LOC"
            elif col == "Ni": name_key = "PMI_NAME_NI"
            elif col == "Cr": name_key = "PMI_NAME_CR"
            elif col == "Mo": name_key = "PMI_NAME_MO"
            elif col == "Result": name_key = "PMI_NAME_RES"
            else: name_key = None
            
            display_text = self.config.get(name_key, col) if name_key else col
            self.preview_tree.heading(col, text=display_text, anchor='center', command=lambda _c=col: self.sort_by_column(_c, mode="PMI"))
            w = saved_widths.get(col, default_widths.get(col, 80))
            self.preview_tree.column(col, width=w, anchor='center', stretch=False)
        
        # [NEW] Add Scrollbars for full coverage
        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.preview_tree.yview)
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.preview_tree.grid(row=0, column=0, sticky='nsew')
        scroll_y.grid(row=0, column=1, sticky='ns')
        scroll_x.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        self._setup_preview_sidebar(self.preview_tree, container, mode="PMI")
        tree_frame.pack(side="left", fill="both", expand=True)

    def _create_rt_preview_ui(self, parent):
        container = tk.Frame(parent, background="#f9fafb")
        container.pack(fill='both', expand=True)

        # [NEW] File Info Header
        header_info = tk.Frame(container, background="#ffffff", highlightthickness=1, highlightbackground="#e5e7eb")
        header_info.pack(fill='x', pady=(0, 5))
        tk.Label(header_info, textvariable=self.file_info_vars['RT'], background="#ffffff", 
                 foreground="#4b5563", font=("Malgun Gothic", 8, "bold"), padx=10, pady=2).pack(side='left')

        self.rt_display_cols = ["V", "No", "Date", "Drawing No.", "Film Ident. No.", "Film Location", "Acc", "Rej", "Deg", "① Crack", "② IP", "③ LF", "④ Slag", "⑤ Por", "⑥ U/C", "⑦ RUC", "⑧ BT", "⑨ TI", "⑩ CP", "⑪ RC", "⑫ Mis", "⑬ EP", "⑭ SD", "⑮ Oth", "Welder No", "Remarks"]
        saved_widths = self.config.get("RT_COL_WIDTHS", {})
        default_widths = {"V": 40, "No": 50, "Date": 90, "Drawing No.": 300, "Film Ident. No.": 120, "Film Location": 100, "Acc": 40, "Rej": 40, "Deg": 40, "Welder No": 100, "Remarks": 120}

        # Inner frame for horizontal/vertical scroll
        tree_frame = tk.Frame(container, background="#f9fafb")
        self.rt_preview_tree = ttk.Treeview(tree_frame, columns=self.rt_display_cols, show='headings', height=10, selectmode='extended')
        for col in self.rt_preview_tree["columns"]:
            name_key = f"RT_NAME_{col.split('(')[0].replace(' ', '').replace('.', '').replace('①', '').replace('②', '').replace('③', '').replace('④', '').replace('⑤', '').upper()}"
            if col == "Drawing No.": name_key = "RT_NAME_DWG"
            elif col == "Film Ident. No.": name_key = "RT_NAME_JOINT"
            elif col == "Film Location": name_key = "RT_NAME_LOC"
            elif col == "Welder No": name_key = "RT_NAME_WELDER"
            elif col == "Acc": name_key = "RT_NAME_ACC"
            elif col == "Rej": name_key = "RT_NAME_REJ"
            elif col == "① Crack": name_key = "RT_NAME_D1"
            elif col == "② IP": name_key = "RT_NAME_D2"
            elif col == "③ LF": name_key = "RT_NAME_D3"
            elif col == "④ Slag": name_key = "RT_NAME_D4"
            elif col == "⑤ Por": name_key = "RT_NAME_D5"
            elif col == "⑥ U/C": name_key = "RT_NAME_D6"
            elif col == "⑦ RUC": name_key = "RT_NAME_D7"
            elif col == "⑧ BT": name_key = "RT_NAME_D8"
            elif col == "⑨ TI": name_key = "RT_NAME_D9"
            elif col == "⑩ CP": name_key = "RT_NAME_D10"
            elif col == "⑪ RC": name_key = "RT_NAME_D11"
            elif col == "⑫ Mis": name_key = "RT_NAME_D12"
            elif col == "⑬ EP": name_key = "RT_NAME_D13"
            elif col == "⑭ SD": name_key = "RT_NAME_D14"
            elif col == "⑮ Oth": name_key = "RT_NAME_D15"
            else: name_key = None
            
            display_text = self.config.get(name_key, col) if name_key else col
            self.rt_preview_tree.heading(col, text=display_text, anchor='center', command=lambda _c=col: self.sort_by_column(_c, mode="RT"))
            w = saved_widths.get(col, default_widths.get(col, 80))
            self.rt_preview_tree.column(col, width=w, anchor='center', stretch=False)
        
        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.rt_preview_tree.yview)
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.rt_preview_tree.xview)
        self.rt_preview_tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.rt_preview_tree.grid(row=0, column=0, sticky='nsew')
        scroll_y.grid(row=0, column=1, sticky='ns')
        scroll_x.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        self._setup_preview_sidebar(self.rt_preview_tree, container, mode="RT")
        tree_frame.pack(side="left", fill="both", expand=True)

    def _setup_preview_sidebar(self, tree, container, mode):
        # 1. Event Bindings
        self.drag_start_item = None
        
        def _on_tree_press(event, t=tree, m=mode):
            t.focus_set()
            
            # Column Drag-and-Drop Init
            if t.identify_region(event.x, event.y) == "heading":
                # Ensure we are not grabbing the separator
                if t.identify_region(event.x - 3, event.y) == "separator" or t.identify_region(event.x + 3, event.y) == "separator":
                    t.drag_col_id = None
                else:
                    t.drag_col_id = t.identify_column(event.x)
                return
            else:
                t.drag_col_id = None

            item_id = t.identify_row(event.y)
            col_id = t.identify_column(event.x)
            
            # [DYNAMIC] Get consistent module info
            mode_info = self._get_mode_info(m)
            if not mode_info: return
            _, idx_map, data_list, col_keys = mode_info
            
            self.start_col = col_id
            self.end_col = col_id
            self.last_clicked_col = col_id
            
            if not item_id or not col_id: return
            
            try:
                col_idx = int(col_id.replace("#", "")) - 1
                if 0 <= col_idx < len(col_keys):
                    key = col_keys[col_idx]
                    # Checkbox Toggle (V or selected)
                    if key in ["selected", "V"] or col_id == "#1": # Fallback to first col for V
                        if key == "selected" or (m == "PMI" and col_id == "#2"):
                             view_idx = t.index(item_id)
                             if 0 <= view_idx < len(idx_map):
                                 actual_idx = idx_map[view_idx]
                                 data_list[actual_idx]['selected'] = not data_list[actual_idx].get('selected', True)
                                 self.populate_preview(data_list, switch_tab=False, mode=m)
                                 return "break"
                    
                    # RT Specific specialized toggle
                    if m == "RT" and ((key.startswith("D") and key[1:].isdigit()) or (key in ["Acc", "Rej"])):
                        view_idx = t.index(item_id)
                        if 0 <= view_idx < len(idx_map):
                            actual_idx = idx_map[view_idx]
                            old_v = data_list[actual_idx].get(key, "")
                            new_v = "√" if old_v == "" else ""
                            data_list[actual_idx][key] = new_v
                            if key in ["Acc", "Rej"] and new_v == "√":
                                other = "Rej" if key == "Acc" else "Acc"
                                data_list[actual_idx][other] = ""
                            self.populate_preview(data_list, switch_tab=False, mode="RT")
                            return "break"
            except: pass

            self.drag_start_item = item_id
            if not (event.state & 0x0001 or event.state & 0x0004):
                t.selection_set(item_id)

        def _on_tree_drag(event, t=tree):
            if getattr(t, 'drag_col_id', None):
                return
                
            if not self.drag_start_item: return
            curr_item = t.identify_row(event.y)
            drag_col = t.identify_column(event.x)
            if drag_col: self.end_col = drag_col
            
            if not curr_item: return
            all_items = t.get_children('')
            try:
                low = min(all_items.index(self.drag_start_item), all_items.index(curr_item))
                high = max(all_items.index(self.drag_start_item), all_items.index(curr_item))
                t.selection_set(all_items[low:high+1])
            except: pass

        def _on_tree_release(event, t=tree):
            self.drag_start_item = None
            
            if getattr(t, 'drag_col_id', None):
                target_col = t.identify_column(event.x)
                if target_col and target_col != t.drag_col_id:
                    try:
                        curr_display = list(t["displaycolumns"])
                        if curr_display == ["#all"]:
                            curr_display = list(t["columns"])
                            
                        drag_id = t.column(t.drag_col_id, "id")
                        target_id = t.column(target_col, "id")
                        
                        if drag_id in curr_display and target_id in curr_display:
                            curr_display.remove(drag_id)
                            target_idx = curr_display.index(target_id)
                            curr_display.insert(target_idx, drag_id)
                            t["displaycolumns"] = tuple(curr_display)
                    except Exception:
                        pass
                t.drag_col_id = None

        tree.bind("<Button-1>", _on_tree_press)
        tree.bind("<B1-Motion>", _on_tree_drag)
        tree.bind("<ButtonRelease-1>", _on_tree_release)
        tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        tree.bind("<Double-1>", lambda e, m=mode: self.on_tree_double_click(e, m))
        tree.bind("<Button-3>", lambda e, m=mode: self.show_context_menu(e, m))
        tree.bind("<Control-c>", lambda e, m=mode: self.copy_cell(m))
        tree.bind("<Control-C>", lambda e, m=mode: self.copy_cell(m))
        tree.bind("<Control-v>", lambda e, m=mode: self.paste_cell(m))
        tree.bind("<Control-V>", lambda e, m=mode: self.paste_cell(m))
        tree.bind("<Control-a>", lambda e, t=tree: [t.selection_set(t.get_children(''))])
        tree.bind("<Control-A>", lambda e, t=tree: [t.selection_set(t.get_children(''))])
        tree.bind("<Delete>", lambda e, m=mode: self.on_delete_key(e, m))
        tree.bind("<F2>", lambda e, m=mode: self.on_tree_double_click(e, m))

        # 2. Sidebar with SCROLL support (too many elements to fit without scrolling)
        sidebar_outer = tk.Frame(container, background="#f1f5f9", highlightthickness=1, highlightbackground="#e2e8f0", width=160)
        sidebar_outer.pack(side='left', fill='y', padx=(0, 5))
        sidebar_outer.pack_propagate(False)
        
        # --- STICKY TOP FILTERS ---
        sticky_top_frame = tk.Frame(sidebar_outer, background="#f1f5f9", padx=5, pady=3)
        sticky_top_frame.pack(side='top', fill='x')

        ttk.Label(sticky_top_frame, text="날짜 필터", font=("Malgun Gothic", 8, "bold")).pack(pady=(2, 1))
        date_frame = tk.Frame(sticky_top_frame, background="#f1f5f9")
        date_frame.pack(fill='x', pady=2)
        
        listbox = tk.Listbox(date_frame, selectmode='single', height=5, exportselection=False, font=("Malgun Gothic", 8))
        listbox.pack(side='left', fill='both', expand=True)
        sb = ttk.Scrollbar(date_frame, orient='vertical', command=listbox.yview)
        sb.pack(side='right', fill='y')
        listbox.config(yscrollcommand=sb.set)
        
        if mode == "RT": self.rt_date_listbox = listbox
        elif mode == "PT": self.pt_date_listbox = listbox
        elif mode == "PAUT": self.paut_date_listbox = listbox
        else: self.date_listbox = listbox
        
        def _toggle_date(event, lb=listbox):
            idx = lb.nearest(event.y)
            if idx >= 0:
                val = lb.get(idx)
                if val.startswith("[v]"):
                    lb.delete(idx); lb.insert(idx, val.replace("[v] ", "[ ] "))
                elif val.startswith("[ ]"):
                    lb.delete(idx); lb.insert(idx, val.replace("[ ] ", "[v] "))
                lb.selection_clear(0, tk.END)
        
        listbox.bind("<ButtonRelease-1>", _toggle_date)
        
        def _apply_date_filter(m=mode, lb=listbox):
            sel_dates = [lb.get(i).replace("[v] ", "").replace("[ ] ", "") for i in range(lb.size()) if lb.get(i).startswith("[v]")]
            data = self.rt_extracted_data if m == "RT" else self.extracted_data
            for item in data:
                item['date_filtered'] = (item.get('Date', '') in sel_dates)
            
            if m == "PMI":
                self.update_pmi_loc_listbox()
                
            self.populate_preview(data, switch_tab=False, mode=m)
            
        ttk.Button(sticky_top_frame, text="날짜 적용", command=_apply_date_filter).pack(fill='x', pady=(0, 5))
        
        # [NEW] Multi-Select Test Location Filter (PMI Only)
        if mode == "PMI":
            ttk.Label(sticky_top_frame, text="지점(Loc) 필터", font=("Malgun Gothic", 8, "bold"), background="#f1f5f9").pack(pady=(2, 1))
            loc_frame = tk.Frame(sticky_top_frame, background="#f1f5f9")
            loc_frame.pack(fill='x', pady=2)
            
            self.pmi_loc_listbox = tk.Listbox(loc_frame, selectmode='single', height=4, exportselection=False, font=("Malgun Gothic", 8))
            self.pmi_loc_listbox.pack(side='left', fill='both', expand=True)
            sb_loc = ttk.Scrollbar(loc_frame, orient='vertical', command=self.pmi_loc_listbox.yview)
            sb_loc.pack(side='right', fill='y')
            self.pmi_loc_listbox.config(yscrollcommand=sb_loc.set)
            
            def _toggle_loc(event):
                idx = self.pmi_loc_listbox.nearest(event.y)
                if idx >= 0:
                    val = self.pmi_loc_listbox.get(idx)
                    if val.startswith("[v]"):
                        self.pmi_loc_listbox.delete(idx); self.pmi_loc_listbox.insert(idx, val.replace("[v] ", "[ ] "))
                    elif val.startswith("[ ]"):
                        self.pmi_loc_listbox.delete(idx); self.pmi_loc_listbox.insert(idx, val.replace("[ ] ", "[v] "))
                    self.pmi_loc_listbox.selection_clear(0, tk.END)
            
            self.pmi_loc_listbox.bind("<ButtonRelease-1>", _toggle_loc)
            
            def _apply_loc_filter():
                self.populate_preview(self.extracted_data, switch_tab=False, mode="PMI")
            
            ttk.Button(sticky_top_frame, text="지점 적용", command=_apply_loc_filter).pack(fill='x', pady=(0, 5))

        # [NEW] Sticky 'Show Selected Only' Checkbox
        tk.Checkbutton(sticky_top_frame, text="선택만", variable=self.show_selected_only, 
                       command=lambda: self.populate_preview(self.rt_extracted_data if mode=="RT" else self.extracted_data, switch_tab=False, mode=mode),
                       background="#f1f5f9", font=("Malgun Gothic", 8)).pack(pady=(0, 3), anchor='w')

        # --- SCROLLABLE BUTTONS AREA ---
        scroll_wrapper = tk.Frame(sidebar_outer, background="#f1f5f9")
        scroll_wrapper.pack(side='top', fill='both', expand=True)

        sidebar_canvas = tk.Canvas(scroll_wrapper, background="#f1f5f9", highlightthickness=0)
        sidebar_sb = ttk.Scrollbar(scroll_wrapper, orient="vertical", command=sidebar_canvas.yview)
        sidebar = tk.Frame(sidebar_canvas, background="#f1f5f9", padx=5, pady=3)
        
        sidebar.bind("<Configure>", lambda e: sidebar_canvas.configure(scrollregion=sidebar_canvas.bbox("all")))
        # Fixed width for inner sidebar relative to outer width (160)
        # Scrollbar takes about ~15px, let's make window ~140px
        sidebar_canvas.create_window((0, 0), window=sidebar, anchor="nw", width=140)
        sidebar_canvas.configure(yscrollcommand=sidebar_sb.set)
        
        sidebar_canvas.pack(side='left', fill='both', expand=True)
        sidebar_sb.pack(side='right', fill='y')
        
        # Mousewheel scroll for sidebar
        def _sidebar_scroll(event):
            # Bulletproof bounds checking to ensure no overscroll detach
            top, bottom = sidebar_canvas.yview()
            direction = int(-1*(event.delta/120))
            
            # If everything fits, lock it
            if top <= 0.0 and bottom >= 1.0:
                sidebar_canvas.yview_moveto(0)
                return "break"
                
            if direction < 0 and top <= 0.0:
                sidebar_canvas.yview_moveto(0)
                return "break"
                
            if direction > 0 and bottom >= 1.0:
                sidebar_canvas.yview_moveto(1.0)
                return "break"
                
            # Perform scroll
            sidebar_canvas.yview_scroll(direction, "units")
            
            # Post-scroll enforcement for overshoot
            new_top, new_bottom = sidebar_canvas.yview()
            if direction < 0 and new_top <= 0.0:
                sidebar_canvas.yview_moveto(0)
            elif direction > 0 and new_bottom >= 1.0:
                sidebar_canvas.yview_moveto(1.0)
                
            return "break"
            
        sidebar_canvas.bind("<MouseWheel>", _sidebar_scroll)
        sidebar.bind("<MouseWheel>", _sidebar_scroll)
        def _bind_mousewheel_recursive(widget):
            widget.bind("<MouseWheel>", _sidebar_scroll)
            for child in widget.winfo_children():
                _bind_mousewheel_recursive(child)
        sidebar.bind("<Configure>", lambda e: _bind_mousewheel_recursive(sidebar), add="+")

        ttk.Button(sidebar, text="전체 선택", width=12, command=lambda: self.select_all(mode)).pack(pady=2, fill='x')
        ttk.Button(sidebar, text="선택 해제", width=12, command=lambda: self.deselect_all(mode)).pack(pady=2, fill='x')
        
        if mode == "PMI":
            ttk.Button(sidebar, text="미달만 선택", width=12, command=lambda: self.select_deficient_items()).pack(pady=2, fill='x')
        
        ttk.Button(sidebar, text="컬럼 관리", width=12, command=lambda: self.manage_columns(mode)).pack(pady=2, fill='x')
        
        if mode == "PAUT":
            ttk.Button(sidebar, text="행 추가", width=12, command=lambda: self.add_item(mode)).pack(pady=2, fill='x')
        
        ttk.Button(sidebar, text="병합", width=12, command=lambda: self.merge_selected_iso(mode)).pack(pady=2, fill='x')
        ttk.Button(sidebar, text="일괄 변경", width=12, command=lambda: self.show_bulk_update_dialog(mode)).pack(pady=(2, 3), fill='x')

        tk.Frame(sidebar, height=1, background="#e5e7eb").pack(fill='x', pady=5)
        ttk.Button(sidebar, text="▲ 위로", width=10, command=lambda: self.move_item(-1, mode)).pack(pady=5)
        ttk.Button(sidebar, text="▼ 아래로", width=10, command=lambda: self.move_item(1, mode)).pack(pady=5)
        tk.Frame(sidebar, height=10, background="#f9fafb").pack()
        ttk.Button(sidebar, text="선택 삭제", width=12, command=lambda: self.delete_item(mode)).pack(pady=5)
        ttk.Button(sidebar, text="전체 초기화", width=12, command=lambda: self.clear_all(mode)).pack(pady=(2, 10))

        tk.Frame(sidebar, height=1, background="#e5e7eb").pack(fill='x', pady=5)
        ttk.Button(sidebar, text="💾 현재 내용 저장", width=18, command=lambda: self.save_preview_data(mode)).pack(pady=5)
        ttk.Button(sidebar, text="📂 저장된 내용 열기", width=18, command=lambda: self.load_preview_data(mode)).pack(pady=5)
        ttk.Button(sidebar, text="📊 엑셀 파일로 추출", width=18, command=lambda: self.export_to_excel(mode)).pack(pady=(10, 2))

        # 3. Context Menu Config
        for t in ["evenrow", "oddrow", "group_even", "group_odd", "grouped_even", "grouped_odd", "item_even", "item_odd"]:
            tree.tag_configure(t, background="", foreground="black")
        tree.tag_configure('group_even', font=("Malgun Gothic", 10, "bold"))
        tree.tag_configure('group_odd', font=("Malgun Gothic", 10, "bold"))


    def on_tree_select(self, event):
        """항목 선택 시 체크박스 상태와 동기화 (현재는 드래그 안정성을 위해 드래그 중에는 무시)"""
        pass

    def toggle_item_selection(self, item_id, mode="PMI"):
        """단일 항목의 체크 상태를 토글"""
        tree = self.rt_preview_tree if mode == "RT" else self.preview_tree
        idx_map = self.rt_item_idx_map if mode == "RT" else self.item_idx_map
        data = self.rt_extracted_data if mode == "RT" else self.extracted_data
        
        view_idx = tree.index(item_id)
        if 0 <= view_idx < len(idx_map):
            actual_idx = idx_map[view_idx]
            data[actual_idx]['selected'] = not data[actual_idx].get('selected', True)
            self.populate_preview(data, switch_tab=False, mode=mode)
            
            # 선택 상태 복구
            num_items = len(tree.get_children())
            if view_idx < num_items:
                new_item = tree.get_children()[view_idx]
                tree.selection_set(new_item)
                tree.focus(new_item)

    def toggle_selected_items(self, mode="PMI"):
        """선택된 모든 항목의 체크 상태를 토글"""
        tree = self.rt_preview_tree if mode == "RT" else self.preview_tree
        idx_map = self.rt_item_idx_map if mode == "RT" else self.item_idx_map
        data = self.rt_extracted_data if mode == "RT" else self.extracted_data
        
        selected_ids = tree.selection()
        if not selected_ids: return
        
        for item_id in selected_ids:
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                data[actual_idx]['selected'] = not data[actual_idx].get('selected', True)
        
        self.populate_preview(data, switch_tab=False, mode=mode)
        # 선택 보존
        for sid in selected_ids:
            try: tree.selection_add(sid)
            except: pass

    def on_tree_click_checkbox(self, event):
        """Reserved for future coordinate-based checkbox toggle if needed without blocking drag."""
        pass

    def on_tree_double_click(self, event, mode="PMI"):
        """셀(Cell) 더블클릭 시 엑셀처럼 그 자리에서 바로 수동 편집 가능하게 함"""
        mode_info = self._get_mode_info(mode)
        if not mode_info: return
        tree, idx_map, data_list, col_keys = mode_info

        region = tree.identify_region(event.x, event.y)
        if region != "cell": return
        
        column = tree.identify_column(event.x)
        item_id = tree.identify_row(event.y)
        if not item_id or not column: return
        
        # [DYNAMIC] 컬럼 인덱스를 통해 키 추출
        try:
            col_idx = int(column.replace('#', '')) - 1
            if col_idx < 0 or col_idx >= len(col_keys): return
            key = col_keys[col_idx]
        except: return
        
        # 수정 차단 컬럼 (V, 상태 아이콘 등)
        if key in ["selected", "_status", "V", "ST"]:
            if key in ["selected", "V"]: self.toggle_item_selection(item_id, mode)
            return

        # 현재 값 가져오기
        x, y, w, h = tree.bbox(item_id, column)
        view_idx = tree.index(item_id)
        if not (0 <= view_idx < len(idx_map)): return
        actual_idx = idx_map[view_idx]
        
        curr_val = data_list[actual_idx].get(key, "")
        numeric_keys = ["Th'k(mm)", "Start", "End", "Length(mm)", "Upper", "Lower", "Height(mm)", "a/l", "a/t"]
        if (key in numeric_keys or (key not in ["No", "Joint No.", "Line No.", "Joint", "Loc", "Grade", "Date", "Dwg", "ISO", "ISO/DWG", "Type of Flaw", "Evaluation", "Remarks", "Result", "Location", "Nature", "selected", "_status", "ST", "V"])) and self.to_float(curr_val) > 0:
            curr_val = f"{self.to_float(curr_val):.2f}"
            
        entry = ttk.Entry(tree, font=("Malgun Gothic", 10), exportselection=True)
        entry.insert(0, str(curr_val))
        entry.select_range(0, tk.END)
        entry.place(x=x, y=y, width=w, height=h)
        entry.focus_set()
        
        def finish_edit(e=None):
            if not entry.winfo_exists(): return
            new_val = entry.get().strip()
            
            # [PMI 특화] 함량 소수점 처리 및 판정 재계산
            if mode == "PMI" and key not in ["No", "Joint", "Loc", "Grade", "Date", "Dwg", "_status", "ST", "V", "selected", "order_index"]:
                f_val = self.to_float(new_val)
                data_list[actual_idx][key] = f"{f_val:.2f}" if f_val > 0 else ""
                # 함량 변경 시 등급 판정 재수행
                new_grade = self.check_material_grade(data_list[actual_idx])
                if new_grade: data_list[actual_idx]['Grade'] = new_grade
            elif mode == "PAUT" and key in ["Th'k(mm)", "Height(mm)", "Length(mm)", "Type of Flaw"]:
                data_list[actual_idx][key] = new_val
                # [PAUT] Recalculate Evaluation and Ratios
                t = self.to_float(data_list[actual_idx].get('Th\'k(mm)', 0))
                h = self.to_float(data_list[actual_idx].get('Height(mm)', 0))
                l = self.to_float(data_list[actual_idx].get('Length(mm)', 0))
                d = self.to_float(data_list[actual_idx].get('Upper', 0))
                nat = data_list[actual_idx].get('Type of Flaw', 'Slag')
                eval_mode = self.paut_eval_mode.get()
                
                res, _ = self.evaluate_paut_flaw(t, h, l, d, nat, mode=eval_mode)
                data_list[actual_idx]['Evaluation'] = res
                if l > 0: data_list[actual_idx]['a/l'] = f"{h/l:.3f}"
                if t > 0: data_list[actual_idx]['a/t'] = f"{h/t:.3f}"
            else:
                data_list[actual_idx][key] = new_val

            self.populate_preview(data_list, switch_tab=False, mode=mode)
            self.log(f"📝 {mode} {key} 수정: {new_val}")
            entry.destroy()

        entry.bind("<Return>", finish_edit)
        entry.bind("<FocusOut>", lambda e: finish_edit())
        entry.bind("<Escape>", lambda e: entry.destroy())

    def on_delete_key(self, event, mode="PMI"):
        """Delete 키 선택 시 현재 선택된 셀(들)의 내용을 지움 (엑셀 방식)"""
        mode_info = self._get_mode_info(mode)
        if not mode_info: return
        tree, idx_map, data_list, col_keys = mode_info
        
        selected = tree.selection()
        if not selected: return
        
        # 드래그된 컬럼 범위 계산
        s_col = getattr(self, 'start_col', None)
        e_col = getattr(self, 'end_col', None)
        l_col = getattr(self, 'last_clicked_col', None)
        
        if not s_col or not e_col or s_col == e_col:
            target_col_ids = [l_col or s_col or "#4"]
        else:
            try:
                s_idx = int(s_col.replace('#', '')) - 1
                e_idx = int(e_col.replace('#', '')) - 1
                col_start, col_end = min(s_idx, e_idx), max(s_idx, e_idx)
                target_col_ids = [f"#{i+1}" for i in range(col_start, col_end + 1)]
            except:
                target_col_ids = [l_col or s_col or "#4"]
        
        key_map = {f"#{i+1}": k for i, k in enumerate(col_keys)}
        count = 0
        for item_id in selected:
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                for c_id in target_col_ids:
                    key = key_map.get(c_id)
                    if key and key not in ["selected", "_status"]:
                        data_list[actual_idx][key] = ""
                        count += 1
        
        if count > 0:
            self.populate_preview(data_list, switch_tab=False, mode=mode)
            self.log(f"🧹 {mode} 데이터 {count}건 내용 삭제 완료")

    def show_context_menu(self, event, mode="PMI"):
        """우클릭 컨텍스트 메뉴 표시 및 현재 모드에 따라 명령 구성"""
        tree = self.paut_preview_tree if mode == "PAUT" else (self.rt_preview_tree if mode == "RT" else (self.pt_preview_tree if mode == "PT" else self.preview_tree))
        col = tree.identify_column(event.x)
        self.last_clicked_col = col 
        item_id = tree.identify_row(event.y)
        
        # 선택 상태 보정 (우클릭한 항목으로 선택 변경)
        if item_id:
            if item_id not in tree.selection():
                tree.selection_set(item_id)
            tree.focus(item_id)

        # 메뉴 명령 동적 업데이트
        self.ctx_menu.entryconfigure("선택 항목 체크/해제 토글 (Toggle Check)", command=lambda: self.toggle_selected_items(mode))
        self.ctx_menu.entryconfigure("현재 셀 내용만 복사 (Copy Cell)", command=lambda: self.copy_cell(mode, target="cell"))
        self.ctx_menu.entryconfigure("선택 행 전체 복사 (Copy Row)", command=lambda: self.copy_cell(mode, target="row"))
        self.ctx_menu.entryconfigure("이 컬럼 전체 복사 (Copy Column)", command=lambda: self.copy_cell(mode, target="column"))
        self.ctx_menu.entryconfigure("셀 내용 붙여넣기 (Paste)", command=lambda: self.paste_cell(mode))
        
        self.ctx_menu.entryconfigure("위로 이동 (Move Up)", command=lambda: self.move_item(-1, mode))
        self.ctx_menu.entryconfigure("아래로 이동 (Move Down)", command=lambda: self.move_item(1, mode))
        
        self.ctx_menu.entryconfigure("행 추가 (Add Row)", command=lambda: self.add_item(mode))
        self.ctx_menu.entryconfigure("선택 삭제 (Delete Selected)", command=lambda: self.delete_item(mode))
        
        # PMI 전용/특화 기능 제어
        is_pmi = (mode == "PMI")
        self.ctx_menu.entryconfigure("선택 항목 ISO 병합 (Merge ISO)", command=lambda: self.merge_selected_iso(mode), state="normal" if is_pmi else "disabled")
        self.ctx_menu.entryconfigure("선택 항목 Joint 병합 (Merge Joint)", command=lambda: self.merge_selected_joint(mode), state="normal" if is_pmi else "disabled")
        self.ctx_menu.entryconfigure("선택 항목 ISO 병합 해제 (Ungroup ISO)", command=lambda: self.ungroup_selected_iso(mode), state="normal" if is_pmi else "disabled")
        self.ctx_menu.entryconfigure("선택 항목 Joint 병합 해제 (Ungroup Joint)", command=lambda: self.ungroup_selected_joint(mode), state="normal" if is_pmi else "disabled")
        self.ctx_menu.entryconfigure("선택 항목 일괄 변경 (Bulk Update)", command=lambda: self.show_bulk_update_dialog(mode))

        # 컬럼 설정 기능 (모든 모드 지원)
        self.ctx_menu.entryconfigure("컬럼 설정 (Column Manager)", command=lambda: self.manage_columns(mode))
        self.ctx_menu.entryconfigure("컬럼 설정 (Column Manager)", state="normal")

        self.ctx_menu.post(event.x_root, event.y_root)

    def manage_columns(self, mode="PMI"):
        """미리보기에서 표시할 컬럼을 동적으로 관리하는 다이얼로그 (모든 모드 지원)"""
        if mode == "RT":
            tree, idx_map, data, keys_attr = self.rt_preview_tree, self.rt_item_idx_map, self.rt_extracted_data, "rt_column_keys"
        elif mode == "PT":
            tree, idx_map, data, keys_attr = self.pt_preview_tree, self.pt_item_idx_map, self.pt_extracted_data, "pt_column_keys"
        elif mode == "PAUT":
            tree, idx_map, data, keys_attr = self.paut_preview_tree, self.paut_item_idx_map, self.paut_extracted_data, "paut_column_keys"
        else:
            tree, idx_map, data, keys_attr = self.preview_tree, self.item_idx_map, self.extracted_data, "column_keys"

        dialog = tk.Toplevel(self.root)
        dialog.title(f"{mode} 컬럼 관리")
        dialog.geometry("400x550")
        dialog.transient(self.root)
        dialog.grab_set()
        
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill='both', expand=True)
        
        ttk.Label(main_frame, text=f"{mode} 표시할 컬럼을 선택하세요:", font=("Malgun Gothic", 10, "bold")).pack(pady=(0, 10), anchor='w')
        
        # 엑셀 데이터 또는 기존 정의에서 모든 키 추출
        all_keys = []
        if data:
            all_keys = list(data[0].keys())
        else:
            # 데이터가 없을 경우 현재 정의된 키들을 기본으로 함
            all_keys = list(getattr(self, keys_attr))
            
        # 내부 관리용 키 제외
        exclude = ["selected", "date_filtered", "order_index", "visual_group_joint", "is_merged_iso", "is_merged_joint"]
        display_keys = [k for k in all_keys if k not in exclude]
        
        # 스크롤 가능한 영역
        scroll_frame = ttk.Frame(main_frame)
        scroll_frame.pack(fill='both', expand=True)
        
        canvas = tk.Canvas(scroll_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(scroll_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # RT용 레이블 매핑 정의
        label_map = {}
        if mode == "RT":
            label_map = {
                "Dwg": "Drawing No.", "Joint": "Film Ident. No.", "Loc": "Film Location",
                "Accept": "Acc", "Reject": "Rej", "Grade": "Deg", "Welder": "Welder No",
                "D1": "① Crack", "D2": "② IP", "D3": "③ LF", "D4": "④ Slag", "D5": "⑤ Por",
                "D6": "⑥ U/C", "D7": "⑦ RUC", "D8": "⑧ BT", "D9": "⑨ TI", "D10": "⑩ CP",
                "D11": "⑪ RC", "D12": "⑫ Mis", "D13": "⑬ EP", "D14": "⑭ SD", "D15": "⑮ Oth"
            }
        elif mode == "PAUT":
            label_map = {"ISO": "ISO/DWG", "Location": "Loc"}

        vars_dict = {}
        current_keys = getattr(self, keys_attr)
        for k in display_keys:
            v = tk.BooleanVar(value=(k in current_keys))
            # 매핑된 레이블이 있으면 사용, 없으면 키 원본 사용
            display_name = label_map.get(k, k)
            cb = ttk.Checkbutton(scrollable_frame, text=display_name, variable=v)
            cb.pack(fill='x', anchor='w', pady=5)
            vars_dict[k] = v

        # 컬럼 관리 버튼 영역 (추가/삭제)
        mgr_btn_frame = ttk.Frame(main_frame)
        mgr_btn_frame.pack(fill='x', pady=5)
        
        from tkinter import simpledialog

        def _add_col():
            name = simpledialog.askstring("새 컬럼 추가", "추가할 컬럼 이름을 입력하세요:", parent=dialog)
            if name:
                name = name.strip()
                if not name: return
                if any(k.lower() == name.lower() for k in all_keys):
                    messagebox.showerror("오류", "이미 존재하는 컬럼 이름입니다.", parent=dialog)
                    return
                # 데이터 구조에 반영
                if data:
                    for item in data:
                        if name not in item: item[name] = ""
                # 현재 키 리스트에도 즉시 반영 (데이터가 없을 때를 위해)
                curr = getattr(self, keys_attr)
                if name not in curr: curr.append(name)

                self.log(f"➕ {mode} 새 컬럼 추가됨: {name}")
                dialog.destroy()
                self.manage_columns(mode)

        def _del_col():
            name = simpledialog.askstring("컬럼 삭제", "삭제할 컬럼 이름을 정확히 입력하세요:", parent=dialog)
            if name:
                name = name.strip()
                if not name: return
                essential = ["V", "No", "selected", "ISO", "Joint", "ISO/DWG", "Dwg"]
                if any(k.lower() == name.lower() for k in essential):
                    messagebox.showerror("오류", f"'{name}' 컬럼은 필수 항목이므로 삭제할 수 없습니다.", parent=dialog)
                    return
                
                target_key = next((k for k in display_keys if k.lower() == name.lower()), None)
                if not target_key:
                    messagebox.showerror("오류", f"'{name}' 컬럼을 찾을 수 없습니다.", parent=dialog)
                    return
                
                if messagebox.askyesno("삭제 확인", f"'{target_key}' 컬럼을 데이터에서 완전히 삭제하시겠습니까?", parent=dialog):
                    if data:
                        for item in data:
                            if target_key in item: del item[target_key]
                    curr = getattr(self, keys_attr)
                    if target_key in curr: curr.remove(target_key)

                    self.log(f"🗑️ {mode} 컬럼 삭제됨: {target_key}")
                    dialog.destroy()
                    self.manage_columns(mode)

        ttk.Button(mgr_btn_frame, text="컬럼 추가", command=_add_col).pack(side='left', padx=5, expand=True)
        ttk.Button(mgr_btn_frame, text="컬럼 삭제", command=_del_col).pack(side='left', padx=5, expand=True)
            
        def _apply():
            selected_keys = []
            
            if "_status" in vars_dict and vars_dict["_status"].get():
                selected_keys.append("_status")
            elif "ST" in vars_dict and vars_dict["ST"].get():
                selected_keys.append("ST")
                
            selected_keys.append("selected") # "V" column is required
            
            if "No" in vars_dict and vars_dict["No"].get():
                selected_keys.append("No")
            elif "No" not in selected_keys:
                selected_keys.append("No") # Fallback safety

            for k in display_keys:
                if k not in ["ST", "_status", "No"] and vars_dict[k].get():
                    selected_keys.append(k)
            
            setattr(self, keys_attr, selected_keys)
            
            # Treeview Columns Re-setup
            header_mapping = {}
            current_cols = []
            
            # Fetch existing drag-rearranged visual order for persistence
            old_display_cols = list(tree["displaycolumns"])
            if old_display_cols == ["#all"] or not old_display_cols:
                old_display_cols = list(tree["columns"])
            
            for k in selected_keys:
                # 내부 ID 결정
                col_id = k
                if k == "selected": col_id = "V"
                elif k == "_status": col_id = "ST"
                elif k in ["ISO", "ISO/DWG", "Dwg"]: col_id = ("Dwg" if mode == "RT" else "ISO/DWG")
                elif mode == "PMI":
                    if k == "Joint": col_id = "Joint No"
                    elif k == "Loc": col_id = "Test Location"
                
                current_cols.append(col_id)
                # 표시용 헤더 텍스트 결정
                if k == "Joint" and mode != "RT":
                    header_mapping[col_id] = "Joint No"
                elif mode == "PMI" and k == "Loc":
                    header_mapping[col_id] = "Test Location"
                else:
                    header_mapping[col_id] = label_map.get(k, col_id)

            # Reconstruct sequence preserving old elements, adding new ones at the end
            new_display = [x for x in old_display_cols if x in current_cols]
            for x in current_cols:
                if x not in new_display:
                    new_display.append(x)

            tree["columns"] = tuple(current_cols)
            tree["displaycolumns"] = tuple(new_display)
            
            for col in tree["columns"]:
                # Width heuristics
                w = 120 if any(x in col.upper() for x in ["ISO", "DWG", "JOINT"]) else (40 if col in ["V", "No", "t", "h", "l", "d"] else 80)
                tree.heading(col, text=header_mapping.get(col, col), anchor='center', command=lambda _c=col: self.sort_by_column(_c, mode=mode))
                tree.column(col, width=w, anchor='center', stretch=False)
            
            self.populate_preview(data if data else [], switch_tab=False, mode=mode)
            self.log(f"⚙️ {mode} 미리보기 컬럼 설정 변경 완료")
            dialog.destroy()
            
        ttk.Button(main_frame, text="적용 (Apply)", command=_apply).pack(pady=5)


    def update_date_listbox(self, mode="PMI"):
        if mode == "RT":
            listbox = self.rt_date_listbox
            data = self.rt_extracted_data
        elif mode == "PT":
            listbox = self.pt_date_listbox
            data = self.pt_extracted_data
        else:
            listbox = self.date_listbox
            data = self.extracted_data
        
        if not listbox: return
        
        listbox.delete(0, tk.END)
        # [NEW] 데이터 기반으로 날짜별 선택 상태(date_filtered) 수집
        date_status = {}
        for item in data:
            dt = item.get('Date', 'N/A')
            if dt not in date_status:
                date_status[dt] = item.get('date_filtered', True)
        
        dates = sorted(list(date_status.keys()))
        for d in dates:
            is_v = date_status.get(d, True)
            prefix = "[v]" if is_v else "[ ]"
            listbox.insert(tk.END, f"{prefix} {d}")
            if is_v: listbox.select_set(tk.END)

    def _get_mode_info(self, mode):
        """Helper to get core UI/Data objects for a specific module mode."""
        if mode == "PMI":
            return (self.preview_tree, self.item_idx_map, self.extracted_data, self.column_keys)
        elif mode == "RT":
            return (self.rt_preview_tree, self.rt_item_idx_map, self.rt_extracted_data, self.rt_column_keys)
        elif mode == "PT":
            return (self.pt_preview_tree, self.pt_item_idx_map, self.pt_extracted_data, self.pt_column_keys)
        elif mode == "PAUT":
            return (self.paut_preview_tree, self.paut_item_idx_map, self.paut_extracted_data, self.paut_column_keys)
        return None

    def copy_cell(self, mode="PMI", target="smart"):
        """선택된 영역의 내용을 엑셀 형식(\t)으로 클립보드에 복사"""
        try:
            mode_info = self._get_mode_info(mode)
            if not mode_info: return
            tree, idx_map, data_list, col_keys = mode_info

            selected = list(tree.selection())
            if not selected: return
            selected.sort(key=lambda x: tree.index(x))

            # [REFINED] 드래그 영역 계산
            s_col = getattr(self, 'start_col', None)
            e_col = getattr(self, 'end_col', None)
            l_col = getattr(self, 'last_clicked_col', None)

            if target == "cell":
                target_col_ids = [l_col or s_col or "#4"]
            elif not s_col or not e_col or s_col == e_col:
                target_col_ids = [l_col or s_col or "#4"]
            else:
                try:
                    s_idx = int(s_col.replace('#', '')) - 1
                    e_idx = int(e_col.replace('#', '')) - 1
                    col_start, col_end = min(s_idx, e_idx), max(s_idx, e_idx)
                    target_col_ids = [f"#{i+1}" for i in range(col_start, col_end + 1)]
                except:
                    target_col_ids = [l_col or s_col or "#4"]

            key_map = {f"#{i+1}": k for i, k in enumerate(col_keys)}
            output = []

            for item_id in selected:
                view_idx = tree.index(item_id)
                if not (0 <= view_idx < len(idx_map)): continue
                actual_idx = idx_map[view_idx]
                item = data_list[actual_idx]
                
                row_vals = []
                for c_id in target_col_ids:
                    key = key_map.get(c_id)
                    if key == "selected":
                        row_vals.append("●" if item.get('selected', True) else "○")
                    elif key == "_status":
                        is_def = False
                        if mode == "PMI" and hasattr(self, 'element_filters'):
                            for f in self.element_filters:
                                fk = f['key'].get().strip()
                                if not fk: continue
                                v = self.to_float(self._get_val_ci(item, fk))
                                fm, fx = self.to_float(f['min'].get()), self.to_float(f['max'].get())
                                if (fm > 0 and v < fm) or (fx > 0 and v > fx):
                                    is_def = True; break
                        row_vals.append("⚠️" if is_def else "✅")
                    else:
                        val = item.get(key, "")
                        if mode == "PMI" and key not in ["No", "Joint", "Loc", "Grade", "Date", "Dwg", "_status", "ST", "V", "selected", "order_index"]:
                            f_val = self.to_float(val)
                            val = f"{f_val:.2f}" if f_val > 0 else ""
                        row_vals.append(str(val))
                
                output.append("\t".join(row_vals))

            self.root.clipboard_clear()
            self.root.clipboard_append("\n".join(output))
            self.log(f"📋 {mode} 복사 완료: {len(selected)}행 x {len(target_col_ids)}열")
        except Exception as e:
            self.log(f"❌ 복사 오류: {e}")

    def paste_cell(self, mode="PMI"):
        """클립보드 내용을 선택된 영역에 스마트하게(Replicate 지원) 붙여넣기"""
        try:
            mode_info = self._get_mode_info(mode)
            if not mode_info: return
            tree, idx_map, data_list, col_keys = mode_info
            
            try: 
                clipboard_val = self.root.clipboard_get()
                if not clipboard_val: return
                # 엑셀 형식(\t, \n) 파싱
                paste_rows = [line.split("\t") for line in clipboard_val.strip().split("\n")]
            except Exception as e:
                self.log(f"⚠️ 클립보드 데이터를 가져올 수 없습니다: {e}")
                return
            
            selected = list(tree.selection())
            if not selected:
                messagebox.showwarning("붙여넣기 오류", "붙여넣을 셀을 선택해주세요.")
                return
            selected.sort(key=lambda x: tree.index(x))
            
            # [SMART AREA] 드래그된 컬럼 범위를 계산
            s_col = getattr(self, 'start_col', None)
            e_col = getattr(self, 'end_col', None)
            l_col = getattr(self, 'last_clicked_col', None)
            
            if not s_col or not e_col or s_col == e_col:
                target_col = l_col or s_col or "#4"
                try:
                    s_idx = int(target_col.replace('#', '')) - 1
                    col_start = s_idx
                    # 단일 셀 선택 시 클립보드 너비만큼 자동 확장
                    max_cols_in_paste = len(paste_rows[0]) if paste_rows else 1
                    col_end = min(len(col_keys) - 1, col_start + max_cols_in_paste - 1)
                except:
                    col_start, col_end = 0, len(col_keys) - 1
            else:
                try:
                    s_idx = int(s_col.replace('#', '')) - 1
                    e_idx = int(e_col.replace('#', '')) - 1
                    col_start, col_end = min(s_idx, e_idx), max(s_idx, e_idx)
                except:
                    col_start, col_end = 0, len(col_keys) - 1
            
            target_col_ids = [f"#{i+1}" for i in range(col_start, col_end + 1)]
            key_map = {f"#{i+1}": k for i, k in enumerate(col_keys)}
            
            # [EXCEL-LIKE REPLICATION]
            # 한 행만 선택했는데 클립보드가 여러 행이면 -> 자동 드래그 (Fill Down)
            all_children = tree.get_children('')
            if len(selected) == 1 and len(paste_rows) > 1:
                curr_idx = all_children.index(selected[0])
                for i in range(1, len(paste_rows)):
                    if curr_idx + i < len(all_children):
                        selected.append(all_children[curr_idx + i])
            
            for r_idx, item_id in enumerate(selected):
                view_idx = tree.index(item_id)
                if not (0 <= view_idx < len(idx_map)): continue
                actual_idx = idx_map[view_idx]
                
                # 행 복제(Replicate): 클립보드 행보다 선택된 행이 많으면 순환하며 붙여넣음
                source_row = paste_rows[r_idx % len(paste_rows)]
                
                for c_idx, c_id in enumerate(target_col_ids):
                    # 열 복제(Replicate): 클립보드 열보다 선택된 열이 많으면 순환하며 붙여넣음
                    cell_val = source_row[c_idx % len(source_row)].strip()
                    key = key_map.get(c_id)
                    
                    if key and key not in ["selected", "_status"]:
                        if mode == "PMI" and key not in ["No", "Joint", "Loc", "Grade", "Date", "Dwg", "_status", "ST", "V", "selected", "order_index"]:
                            f_val = self.to_float(cell_val)
                            data_list[actual_idx][key] = f"{f_val:.2f}" if f_val > 0 else ""
                        else:
                            data_list[actual_idx][key] = cell_val
            
            self.populate_preview(data_list, switch_tab=False, mode=mode)
            self.log(f"📋 {mode} 붙여넣기 완료: {len(selected)}행 x {len(target_col_ids)}열")
        except Exception as e:
            self.log(f"❌ 붙여넣기 오류: {e}")
            messagebox.showerror("붙여넣기 실패", f"데이터를 붙여넣는 중 오류가 발생했습니다:\n{e}")

    def merge_selected_iso(self, mode="PMI"):
        """선택된 항목들의 ISO 번호를 첫 번째 항목의 것으로 통일 (병합 효과)"""
        mode_info = self._get_mode_info(mode)
        if not mode_info: return
        tree, idx_map, data_list, _ = mode_info
        
        selected = tree.selection()
        if len(selected) < 2:
            messagebox.showinfo("알림", "병합할 항목을 2개 이상 선택해주세요.")
            return
        
        selected = sorted(list(selected), key=lambda x: tree.index(x))
        view_idx_0 = tree.index(selected[0])
        actual_idx_0 = idx_map[view_idx_0] if 0 <= view_idx_0 < len(idx_map) else None
        
        k_iso = "ISO" if mode == "PAUT" else "Dwg"
        if actual_idx_0 is not None:
            first_iso = data_list[actual_idx_0].get(k_iso, '')
        else:
            first_iso = ""
        
        for i, item_id in enumerate(selected):
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                data_list[actual_idx][k_iso] = first_iso
                if i > 0: data_list[actual_idx]['is_merged_iso'] = True
                else: data_list[actual_idx].pop('is_merged_iso', None)
        
        self.populate_preview(data_list, switch_tab=False, mode=mode)
        self.log(f"🔗 {mode} {len(selected)}개 항목 ISO 병합 완료: {first_iso}")

    def merge_selected_joint(self, mode="PMI"):
        """선택된 항목들의 Joint No를 첫 번째 항목의 것으로 통일 (데이터 완전 변경)"""
        mode_info = self._get_mode_info(mode)
        if not mode_info: return
        tree, idx_map, data_list, _ = mode_info
        
        selected = tree.selection()
        if len(selected) < 2:
            messagebox.showinfo("알림", "병합할 항목을 2개 이상 선택해주세요.")
            return
        
        selected = sorted(list(selected), key=lambda x: tree.index(x))
        view_idx_0 = tree.index(selected[0])
        actual_idx_0 = idx_map[view_idx_0] if 0 <= view_idx_0 < len(idx_map) else None
        
        if actual_idx_0 is not None:
            first_joint = data_list[actual_idx_0].get('Joint', '')
        else:
            first_joint = ""
        
        for i, item_id in enumerate(selected):
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                data_list[actual_idx]['Joint'] = first_joint
                if i > 0: data_list[actual_idx]['is_merged_joint'] = True
                else: data_list[actual_idx].pop('is_merged_joint', None)
        
        self.populate_preview(data_list, switch_tab=False, mode=mode)
        self.log(f"🔗 {mode} {len(selected)}개 항목 Joint 병합 완료: {first_joint}")

    def group_selected_joint(self, mode="PMI"):
        """선택된 항목들을 시각적으로만 같은 Joint로 취급하여 묶음 (데이터 유지)"""
        tree = self.rt_preview_tree if mode == "RT" else self.preview_tree
        idx_map = self.rt_item_idx_map if mode == "RT" else self.item_idx_map
        data = self.rt_extracted_data if mode == "RT" else self.extracted_data
        
        selected = tree.selection()
        if len(selected) < 2:
            messagebox.showinfo("알림", "그룹화할 항목을 2개 이상 선택해주세요.")
            return
            
        selected = sorted(list(selected), key=lambda x: tree.index(x))
            
        view_idx_0 = tree.index(selected[0])
        actual_idx_0 = idx_map[view_idx_0] if 0 <= view_idx_0 < len(idx_map) else None
        
        if actual_idx_0 is not None:
            group_val = data[actual_idx_0].get('Joint', '')
        else:
            group_val = ""
            
        for i, item_id in enumerate(selected):
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                data[actual_idx]['visual_group_joint'] = group_val
                if i > 0: data[actual_idx]['is_merged_joint'] = True
                else: data[actual_idx].pop('is_merged_joint', None)
                
        self.populate_preview(data, switch_tab=False, mode=mode)
        self.log(f"🔗 {mode} {len(selected)}개 항목 Joint 시각적 그룹화 완료")

    def ungroup_selected_iso(self, mode="PMI"):
        """선택된 항목들의 ISO 시각적 그룹화 연결 끊기"""
        tree = self.rt_preview_tree if mode == "RT" else self.preview_tree
        idx_map = self.rt_item_idx_map if mode == "RT" else self.item_idx_map
        data = self.rt_extracted_data if mode == "RT" else self.extracted_data
        
        selected = tree.selection()
        if not selected: return
            
        for item_id in selected:
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                data[actual_idx].pop('visual_group_iso', None)
                data[actual_idx].pop('is_merged_iso', None)
                
        self.populate_preview(data, switch_tab=False, mode=mode)
        self.log(f"✂️ {mode} {len(selected)}개 항목 ISO 병합 해제 완료")

    def ungroup_selected_iso(self, mode="PMI"):
        """선택된 항목들의 시각적 ISO 병합 해제"""
        tree = self.rt_preview_tree if mode == "RT" else self.preview_tree
        idx_map = self.rt_item_idx_map if mode == "RT" else self.item_idx_map
        data = self.rt_extracted_data if mode == "RT" else self.extracted_data
        
        selected = tree.selection()
        if not selected: return
            
        for item_id in selected:
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                data[actual_idx].pop('visual_group_iso', None)
                data[actual_idx].pop('is_merged_iso', None)
                
        self.populate_preview(data, switch_tab=False, mode=mode)
        self.log(f"✂️ {mode} {len(selected)}개 항목 ISO 병합 해제 완료")

    def ungroup_selected_joint(self, mode="PMI"):
        """선택된 항목들의 Joint 시각적 그룹화 연결 끊기"""
        tree = self.rt_preview_tree if mode == "RT" else self.preview_tree
        idx_map = self.rt_item_idx_map if mode == "RT" else self.item_idx_map
        data = self.rt_extracted_data if mode == "RT" else self.extracted_data
        
        selected = tree.selection()
        if not selected: return
            
        for item_id in selected:
            view_idx = tree.index(item_id)
            if 0 <= view_idx < len(idx_map):
                actual_idx = idx_map[view_idx]
                data[actual_idx].pop('visual_group_joint', None)
                data[actual_idx].pop('is_merged_joint', None)
                
        self.populate_preview(data, switch_tab=False, mode=mode)
        self.log(f"✂️ {mode} {len(selected)}개 항목 Joint 병합 해제 완료")

    def select_all(self, mode="PMI"):
        """모든 항목 체크"""
        if mode == "RT": data = self.rt_extracted_data
        elif mode == "PT": data = self.pt_extracted_data
        elif mode == "PAUT": data = self.paut_extracted_data
        else: data = self.extracted_data
        
        for item in data:
            if item.get('date_filtered', True):
                item['selected'] = True
        self.populate_preview(data, switch_tab=False, mode=mode)

    def deselect_all(self, mode="PMI"):
        """모든 항목 체크 해제"""
        if mode == "RT": data = self.rt_extracted_data
        elif mode == "PT": data = self.pt_extracted_data
        elif mode == "PAUT": data = self.paut_extracted_data
        else: data = self.extracted_data
        
        for item in data:
            if item.get('date_filtered', True):
                item['selected'] = False
        self.populate_preview(data, switch_tab=False, mode=mode)

    def select_deficient_items(self):
        """PMI: 함량 미달인 항목만 자동으로 선택(체크)"""
        if not self.extracted_data or not hasattr(self, 'element_filters'): return
        
        count = 0
        for item in self.extracted_data:
            is_deficient = False
            for f_item in self.element_filters:
                f_key = f_item['key'].get().strip()
                if not f_key: continue
                f_min = round(self.to_float(f_item['min'].get()), 2)
                f_max = round(self.to_float(f_item['max'].get()), 2)
                
                val = round(self.to_float(self._get_val_ci(item, f_key)), 2)
                if (f_min > 0 and val < f_min) or (f_max > 0 and val > f_max):
                    is_deficient = True
                    break
            
            item['selected'] = is_deficient
            if is_deficient: count += 1
            
        self.populate_preview(self.extracted_data, switch_tab=False, mode="PMI")
        self.log(f"✅ 미달 항목 {count}건을 선택했습니다.")

    def reset_pmi_filters(self, refresh=True, clear_search=True):
        """PMI 검색 및 함량 필터 초기화"""
        if clear_search:
            self.pmi_search_loc.set("")
            # 사이드바 지점 필터도 초기화
            if hasattr(self, 'pmi_loc_listbox'):
                for i in range(self.pmi_loc_listbox.size()):
                    val = self.pmi_loc_listbox.get(i)
                    self.pmi_loc_listbox.delete(i)
                    self.pmi_loc_listbox.insert(i, val.replace("[v] ", "[ ] "))
        
        self.pmi_show_deficiency_only.set(False)
        if hasattr(self, 'element_filters'):
            for f in self.element_filters:
                f['min'].set("")
                f['max'].set("")
        if refresh:
            self.populate_preview(self.extracted_data, switch_tab=False, mode="PMI")
        self.log("🧹 PMI 필터가 초기화되었습니다.")

    def update_pmi_loc_listbox(self):
        """현재 데이터의 Test Location 목록을 추출하여 드롭다운 갱신 (날짜 필터 적용된 항목만)"""
        if not hasattr(self, 'pmi_loc_listbox') or not self.extracted_data: return
        
        # 기존 선택 세트 백업
        selected_locs = set()
        for i in range(self.pmi_loc_listbox.size()):
            val = self.pmi_loc_listbox.get(i)
            if val.startswith("[v] "):
                selected_locs.add(val.replace("[v] ", ""))
        
        # 새 목록 추출
        current_locs = sorted(list(set(str(item.get('Loc', '')).strip() for item in self.extracted_data if item.get('Loc') and item.get('date_filtered', True))))
        
        self.pmi_loc_listbox.delete(0, tk.END)
        for loc in current_locs:
            prefix = "[v] " if loc in selected_locs else "[ ] "
            self.pmi_loc_listbox.insert(tk.END, prefix + loc)

    def add_item(self, mode="PMI"):
        """새 데이터 행 추가"""
        if mode == "RT": data = self.rt_extracted_data
        elif mode == "PT": data = self.pt_extracted_data
        elif mode == "PAUT": data = self.paut_extracted_data
        else: data = self.extracted_data
        
        new_item = {'selected': True, 'date_filtered': True, 'order_index': len(data)}
        # 기본 키 초기화
        keys = self.rt_column_keys if mode == "RT" else (self.pt_column_keys if mode == "PT" else (self.paut_column_keys if mode == "PAUT" else self.column_keys))
        for k in keys:
            if k not in new_item: new_item[k] = ""
            
        data.append(new_item)
        self.populate_preview(data, switch_tab=False, mode=mode)
        self.log(f"➕ {mode} 새 데이터 항목 추가 완료")

    def move_item(self, direction, mode="PMI"):
        """선택된 아이템의 순서를 위/아래로 이동 (필터 상태 대응)"""
        # direction이 문자열인 경우 처리 ("up" -> -1, "down" -> 1)
        if isinstance(direction, str):
            direction = -1 if direction.lower() == "up" else 1
        
        mode_info = self._get_mode_info(mode)
        if not mode_info: return
        tree, idx_map, data, _ = mode_info
        
        selected_ids = tree.selection()
        if not selected_ids: return
        
        view_indices = sorted([tree.index(sid) for sid in selected_ids])
        
        if direction == -1: # 위로
            if view_indices[0] == 0: return
        else: # 아래로
            if view_indices[-1] == len(tree.get_children()) - 1: return
            
        # 데이터 리스트(data)에서 위치 교환
        # Treeview 인덱스를 item_idx_map을 통해 실제 데이터 인덱스로 변환
        actual_indices = [idx_map[vi] for vi in view_indices]
        
        # 이동 방향에 따라 처리 순서 결정
        if direction == -1: # 위로
            # 가장 위 아이템의 위쪽 아이템 찾기 (Treeview상에서)
            target_view_idx = view_indices[0] - 1
            target_actual_idx = idx_map[target_view_idx]
            
            # 선택된 아이템들을 한 칸씩 위로 (target_actual_idx 앞으로)
            items_to_move = [data.pop(ai) for ai in reversed(actual_indices)]
            # target_actual_idx가 팝업으로 인해 밀렸을 수 있으므로 다시 계산
            # 하지만 팝업 시 뒤쪽부터 뺐으므로 target_actual_idx 위치는 유지됨
            for item in items_to_move:
                data.insert(target_actual_idx, item)
        else: # 아래로
            # 가장 아래 아이템의 아래쪽 아이템 찾기
            target_view_idx = view_indices[-1] + 1
            target_actual_idx = idx_map[target_view_idx]
            
            # 선택된 아이템들을 한 칸씩 아래로 (target_actual_idx 뒤로)
            items_to_move = [data.pop(ai) for ai in actual_indices]
            # 팝업으로 인해 인덱스가 당겨졌으므로 target_actual_idx에서 len(items_to_move) 만큼 보정하거나 단순히 insert
            insert_pos = target_actual_idx - len(items_to_move) + 1
            for item in reversed(items_to_move):
                data.insert(insert_pos, item)
                
        self.populate_preview(data, switch_tab=False, mode=mode)
        
        # 선택 상태 복원 (Treeview 아이템이 다시 생성되므로 인덱스로 재선택)
        new_children = tree.get_children()
        for vi in view_indices:
            new_pos = max(0, min(len(new_children)-1, vi + direction))
            tree.selection_add(new_children[new_pos])

    def delete_item(self, mode="PMI"):
        """선택된 아이템 삭제"""
        if mode == "RT":
            tree, idx_map, data = self.rt_preview_tree, self.rt_item_idx_map, self.rt_extracted_data
        elif mode == "PT":
            tree, idx_map, data = self.pt_preview_tree, self.pt_item_idx_map, self.pt_extracted_data
        elif mode == "PAUT":
            tree, idx_map, data = self.paut_preview_tree, self.paut_item_idx_map, self.paut_extracted_data
        else:
            tree, idx_map, data = self.preview_tree, self.item_idx_map, self.extracted_data
        
        selected_ids = tree.selection()
        if not selected_ids: return
        
        if not messagebox.askyesno("삭제 확인", f"선택한 {len(selected_ids)}개 항목을 정말 삭제하시겠습니까?"):
            return
            
        view_indices = sorted([tree.index(sid) for sid in selected_ids], reverse=True)
        for vi in view_indices:
            actual_idx = idx_map[vi]
            data.pop(actual_idx)
            
        self.populate_preview(data, switch_tab=False, mode=mode)
        self.update_date_listbox(mode)
        self.log(f"🗑️ {mode} {len(selected_ids)}개 항목 삭제 완료")

    def save_preview_data(self, mode="PMI"):
        """현재 추출/편집된 데이터를 JSON 파일로 저장"""
        if mode == "RT": data = self.rt_extracted_data
        elif mode == "PT": data = self.pt_extracted_data
        elif mode == "PAUT": data = self.paut_extracted_data
        else: data = self.extracted_data
        if not data:
            messagebox.showwarning("알림", f"{mode} 저장할 데이터가 없습니다.")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON Files", "*.json")],
            initialfile=f"{mode}_Data_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            title=f"{mode} 미리보기 데이터 저장"
        )
        
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=4)
                self.log(f"💾 {mode} 데이터 저장 완료: {os.path.basename(file_path)}")
                messagebox.showinfo("완료", "데이터가 성공적으로 저장되었습니다.")
            except Exception as e:
                self.log(f"❌ {mode} 저장 오류: {e}")
                messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다: {e}")

    def load_preview_data(self, mode="PMI"):
        """저장된 JSON 데이터를 불러와서 목록에 채움"""
        file_path = filedialog.askopenfilename(
            filetypes=[("JSON Files", "*.json")],
            title=f"{mode} 미리보기 데이터 불러오기"
        )
        
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    loaded_data = json.load(f)
                    
                if not isinstance(loaded_data, list):
                    raise ValueError("올바른 데이터 형식이 아닙니다.")
                
                data = self.rt_extracted_data if mode == "RT" else self.extracted_data
                
                if data and not messagebox.askyesno("확인", f"현재 작업 중인 {mode} 데이터가 있습니다. 불러온 데이터로 덮어쓰시겠습니까?\n(아니오를 선택하면 기존 데이터에 추가됩니다.)"):
                    data.extend(loaded_data)
                else:
                    if mode == "RT": self.rt_extracted_data = loaded_data
                    else: self.extracted_data = loaded_data
                    data = loaded_data
            
                # [FIX] 필수 속성 보정
                for i, item in enumerate(data):
                    if 'date_filtered' not in item: item['date_filtered'] = True
                    if 'selected' not in item: item['selected'] = True
                    if 'order_index' not in item: item['order_index'] = i
                    
                self.update_date_listbox(mode)
                self.populate_preview(data, mode=mode)
                self.log(f"📂 {mode} 데이터 불러오기 완료: {os.path.basename(file_path)}")
                messagebox.showinfo("완료", f"데이터를 성공적으로 불러왔습니다.\n(총 {len(data)} 건)")
            except Exception as e:
                self.log(f"❌ {mode} 불러오기 오류: {e}")
                messagebox.showerror("오류", f"파일을 불러오는 중 오류가 발생했습니다: {e}")

    def clear_all(self, mode="PMI"):
        """모든 데이터 초기화"""
        if messagebox.askyesno("확인", f"모든 {mode} 데이터를 초기화하시겠습니까?"):
            if mode == "RT": self.rt_extracted_data = []
            elif mode == "PT": self.pt_extracted_data = []
            elif mode == "PAUT": self.paut_extracted_data = []
            else: self.extracted_data = []
            self.update_date_listbox(mode)
            self.populate_preview([], mode=mode)
            self.log(f"🧹 모든 {mode} 데이터 초기화 완료")

    def export_to_excel(self, mode="PMI"):
        """현재 미리보기 목록(필터링/선택 반영)을 엑셀 파일로 내보냄 (서식 및 병합 시인성 유지)"""
        mode_info = self._get_mode_info(mode)
        if not mode_info: return
        _, _, data, col_keys = mode_info
        
        if not data:
            messagebox.showwarning("알림", f"{mode} 내보낼 데이터가 없습니다.")
            return

        # 1. 엑셀에 저장할 대상 데이터 필터링
        filter_enabled = self.show_selected_only.get()
        final_list = [d for d in data if d.get('date_filtered', True) and (not filter_enabled or d.get('selected', True))]
        
        if not final_list:
            messagebox.showinfo("알림", f"{mode} 현재 조건에 맞는 데이터가 없습니다.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"{mode}_Export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            title=f"{mode} 서식 있는 엑셀 파일로 내보내기"
        )
        if not file_path: return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
            import openpyxl.utils

            wb = Workbook()
            ws = wb.active
            ws.title = f"{mode}_Preview"

            # [REFINED] Export with consistent headers from mode_info
            if mode == "PMI":
                # PMI는 특수하게 ST, V 헤더 추가
                headers = ["ST", "V", "No", "Date", "ISO/DWG", "Joint No", "Test Location", "Ni", "Cr", "Mo", "Grade"]
            else:
                # RT/PT/PAUT는 Treeview 컬럼 헤더 그대로 사용
                tree = self._get_mode_info(mode)[0]
                headers = [tree.heading(c)['text'] for c in tree['columns']]
            
            export_rows = []
            for item in final_list:
                row = {}
                if mode == "PMI":
                    # Status Icon
                    is_def = False
                    if hasattr(self, 'element_filters'):
                        for f in self.element_filters:
                            fk = f['key'].get().strip()
                            if not fk: continue
                            v = self.to_float(self._get_val_ci(item, fk))
                            fm, fx = self.to_float(f['min'].get()), self.to_float(f['max'].get())
                            if (fm > 0 and v < fm) or (fx > 0 and v > fx):
                                is_def = True; break
                    
                    row["ST"] = "⚠️" if is_def else "✅"
                    row["V"] = "●" if item.get('selected', True) else "○"
                    row["No"] = item.get('No', '')
                    row["Date"] = item.get('Date', '')
                    row["ISO/DWG"] = item.get('Dwg', '')
                    row["Joint No"] = item.get('Joint', '')
                    row["Test Location"] = item.get('Loc', '')
                    row["Ni"] = f"{self.to_float(item.get('Ni', 0)):.2f}%"
                    row["Cr"] = f"{self.to_float(item.get('Cr', 0)):.2f}%"
                    row["Mo"] = f"{self.to_float(item.get('Mo', 0)):.2f}%"
                    row["Grade"] = item.get('Grade', '')
                else:
                    # Generic mapping for other modes
                    # col_keys: ("_status", "selected", "No", ...)
                    for i, h in enumerate(headers):
                        k = col_keys[i] if i < len(col_keys) else None
                        if k == "_status":
                            row[h] = "✅" 
                        elif k == "selected":
                            row[h] = "●" if item.get('selected', True) else "○"
                        elif k:
                            row[h] = item.get(k, "")
                        else:
                            row[h] = ""
                export_rows.append(row)

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            
            for row_idx, row_data in enumerate(export_rows, 2):
                for col_idx, key in enumerate(headers, 1):
                    val = row_data.get(key, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

            if mode == "PMI":
                column_widths = {1: 5, 2: 5, 3: 8, 4: 12, 5: 25, 6: 15, 7: 15, 8: 8, 9: 8, 10: 8, 11: 15}
            else:
                # RT Column Widths
                column_widths = {1: 5, 2: 8, 3: 12, 4: 25, 5: 15, 6: 10}
                for i in range(7, 22): column_widths[i] = 4 # Defects
                column_widths[22] = 8 # Result
                column_widths[23] = 12 # Welder
                column_widths[24] = 20 # Remarks
                
            for col_idx, width in column_widths.items():
                if col_idx <= len(headers):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = width

            wb.save(file_path)
            self.log(f"📊 {mode} 서식 포함 엑셀 내보내기 완료: {os.path.basename(file_path)}")
            messagebox.showinfo("완료", f"화면 형식과 동일하게 엑셀 파일로 저장했습니다.\n(총 {len(export_rows)} 건)")
        except Exception as e:
            self.log(f"❌ {mode} 엑셀 내보내기 오류: {e}")
            messagebox.showerror("오류", f"엑셀 저장 중 오류가 발생했습니다: {e}")
        finally:
            if 'wb' in locals():
                try: wb.close()
                except: pass

    def sort_by_column(self, col, mode="PMI"):
        """특정 컬럼을 기준으로 데이터 전체 정렬 (문자열/숫자 구분 및 헤더 표시)"""
        # 0. 데이터 대상 선정
        data_targets = {
            "PMI": self.extracted_data,
            "RT": self.rt_extracted_data,
            "PT": self.pt_extracted_data,
            "PAUT": self.paut_extracted_data
        }
        target_list = data_targets.get(mode)
        if not target_list: return
        
        # 1. 정렬 기준 컬럼 매핑 (모든 모드 통합)
        key_map = {
            "selected": "selected", "V": "selected", "No": "No", "Date": "Date", 
            "Dwg": "Dwg", "Joint": "Joint", "Loc": "Loc", "Ni": "Ni", "Cr": "Cr", "Mo": "Mo", "Grade": "Grade",
            "NPS": "NPS", "Thk.": "Thk.", "Material": "Material", "WType": "WType", "Result": "Result",
            "ISO/DWG": "ISO", "t": "t", "h": "h", "l": "l", "d": "d", "Location": "Location", "Nature": "Nature",
            "ST": "_status", "ISO Drawing No.": "Dwg", "Joint No": "Joint", "Test Location": "Loc", "Weld Type": "WType"
        }
        data_key = key_map.get(col)
        if not data_key: return

        # 2. 정렬 방향 결정 (기존 방향과 같으면 토글)
        attr_col = f"{mode.lower()}_sort_col"
        attr_rev = f"{mode.lower()}_sort_rev"
        
        if getattr(self, attr_col, None) == col:
            setattr(self, attr_rev, not getattr(self, attr_rev, False))
        else:
            setattr(self, attr_col, col)
            setattr(self, attr_rev, False)

        sort_rev = getattr(self, attr_rev, False)

        # 3. 데이터 정렬 (지능형 복합 자연어 정렬)
        def get_natural_key(val):
            if val is None: return [(2, "")]
            s_val = str(val).strip().lower()
            if not s_val: return [(2, "")]
            def segment_to_tuple(text):
                if text.isdigit(): return (0, int(text))
                return (1, text)
            return [segment_to_tuple(c) for c in re.split(r'(\d+)', s_val) if c]

        def get_value(item, key):
            if key == "_status" and mode == "PMI":
                is_def = False
                if hasattr(self, 'element_filters'):
                    for f in self.element_filters:
                        fk = f['key'].get().strip()
                        if not fk: continue
                        fm = self.to_float(f['min'].get())
                        fx = self.to_float(f['max'].get())
                        v = self.to_float(self._get_val_ci(item, fk))
                        if (fm > 0 and v < fm) or (fx > 0 and v > fx):
                            is_def = True; break
                return 0 if is_def else 1
            # '_current' 가 넘어오면 클릭된 컬럼(data_key) 값을 반환
            return item.get(data_key if key=="_current" else key, "")

        # 4. 계층적 정렬 수행
        try:
            k_iso = "ISO" if mode == "PAUT" else "Dwg"
            k_joint = "Joint"
            
            sort_key_func = lambda x: (
                get_natural_key(get_value(x, "_current")),
                get_natural_key(get_value(x, k_iso)),
                get_natural_key(get_value(x, k_joint)),
                x.get('order_index', 0)
            )

            target_list.sort(key=sort_key_func, reverse=sort_rev)
        except Exception as e:
            self.log(f"⚠️ {mode} 정렬 오류: {e}")

        # 5. 헤더 텍스트 갱신 (화살표)
        tree_map = {
            "PMI": self.preview_tree, "RT": self.rt_preview_tree,
            "PT": self.pt_preview_tree, "PAUT": self.paut_preview_tree
        }
        target_tree = tree_map.get(mode)
        if target_tree:
            for c in target_tree["columns"]:
                clean_text = c.replace("▲ ", "").replace("▼ ", "")
                # identifier 가 key_map에 기술된 'ST', 'V' 등인 경우와 Dwg 등인 경우 혼재하므로
                # heading text 자체를 조회해서 쓰지 않고 identifier 기반으로 prefix 추가
                prefix = ""
                if c == col:
                    prefix = "▼ " if sort_rev else "▲ "
                target_tree.heading(c, text=prefix + clean_text)

        # 6. 화면 갱신
        self.populate_preview(target_list, switch_tab=False, mode=mode)
        self.log(f"📊 {mode}: '{col}' 기준 {'내림차순' if sort_rev else '오름차순'} 정렬 완료")



    def populate_preview(self, data_list, switch_tab=True, mode="PMI"):
        """추출된 데이터를 미리보기 표에 채움 (필터 반영 및 그룹 색상 적용)"""
        if mode == "RT":
            tree = self.rt_preview_tree
            self.rt_item_idx_map = []
            idx_map = self.rt_item_idx_map
        elif mode == "PT":
            tree = self.pt_preview_tree
            self.pt_item_idx_map = []
            idx_map = self.pt_item_idx_map
        elif mode == "PAUT":
            tree = self.paut_preview_tree
            self.paut_item_idx_map = []
            idx_map = self.paut_item_idx_map
        else:
            tree = self.preview_tree
            self.item_idx_map = []
            idx_map = self.item_idx_map
            # [NEW] Update Loc List
            self.update_pmi_loc_listbox()
            
        # [NEW] Deficiency Detection & Auto-Reset (PMI Only)
            # [REMOVED global check, moved to local loop below]

        for item in tree.get_children():
            tree.delete(item)
        
        filter_enabled = self.show_selected_only.get()
        hidden_by_def_count = 0
        total_matched_search = 0
        local_pmi_deficient_count = 0 # Count within Date/Loc filter
        
        active_locs = set()
        if mode == "PMI" and hasattr(self, 'pmi_loc_listbox'):
            for i in range(self.pmi_loc_listbox.size()):
                val = self.pmi_loc_listbox.get(i)
                if val.startswith("[v] "): active_locs.add(val.replace("[v] ", "").lower())
        
        last_iso = None
        last_joint = None
        current_tag = "group_even"
        
        for idx, item in enumerate(data_list):
            if not item.get('date_filtered', True):
                continue
            
            is_selected = item.get('selected', True)
            if filter_enabled and not is_selected:
                continue
            
            # [NEW] PMI Preview Filters (Multi-Select Test Location & Deficiency)
            selected_locs_names = []
            if mode == "PMI":
                # item Date/Loc filter already passed (Date was line 2608, Loc is here)
                loc_val = str(item.get('Loc', '')).lower()
                if active_locs and loc_val not in active_locs:
                    continue
                
                # Check for deficiency highlighting (AND local count)
                is_deficient = False
                if hasattr(self, 'element_filters'):
                    for f_item in self.element_filters:
                        f_key = f_item['key'].get().strip()
                        if not f_key: continue
                        f_min = round(self.to_float(f_item['min'].get()), 2)
                        f_max = round(self.to_float(f_item['max'].get()), 2)
                        val = round(self.to_float(self._get_val_ci(item, f_key)), 2)
                        if (f_min > 0 and val < f_min) or (f_max > 0 and val > f_max):
                            is_deficient = True; break
                
                if is_deficient: local_pmi_deficient_count += 1
                total_matched_search += 1
                selected_locs_names.append(item.get('Loc', '')) 
                
                if self.pmi_show_deficiency_only.get() and not is_deficient:
                    hidden_by_def_count += 1
                    continue
            else:
                is_deficient = False

            idx_map.append(idx)
            v_mark = "●" if is_selected else "○"
            st_mark = "⚠️" if is_deficient else "✅"
            
            # ISO/DWG 번호가 바뀌면 배경색 태그 교체 (PAUT 'ISO' 대응)
            curr_iso = item.get('Dwg', item.get('ISO', ''))
            norm_iso = self.normalize_iso(curr_iso)
            
            # [NEW] 시각적 병합 (visual_group_joint) 값이 존재하면 해당 값으로 그룹화 계산
            curr_joint = item.get('visual_group_joint', item.get('Joint', ''))
            
            # [FIX] Joint-Aware: Joint가 바뀌었는지도 함께 체크
            is_new_iso = (last_iso is None or self.normalize_iso(last_iso) != norm_iso)
            is_new_joint = (last_joint is None or last_joint != curr_joint)
            
            if is_new_iso:
                current_tag = "group_odd" if current_tag == "group_even" else "group_even"

            # [REFINED] 하이브리드 표시 로직: 블록 시작점 정렬이거나 "새로운 그룹(ISO or Joint)" 시작점일 때 표시
            display_count = len(idx_map) - 1 
            is_block_start = (display_count % 3 == 0)
            
            # ISO나 Joint가 바뀌었거나 3줄 단위 시작이면 표시
            is_show = is_new_iso or is_new_joint or is_block_start

            display_iso = curr_iso if is_show else ""
            display_joint = item.get('Joint', '') if is_show else ""
            
            # [FIX] 명시적으로 병합된 경우라도 '새 그룹 시작'이면 표시 강제
            if item.get('is_merged_iso') and not (is_new_iso or is_new_joint): display_iso = ""
            if item.get('is_merged_joint') and not (is_new_iso or is_new_joint): display_joint = ""

            last_iso = curr_iso
            last_joint = curr_joint
            
            row_vals = []
            if mode == "RT":
                for k in self.rt_column_keys:
                    if k == "selected": row_vals.append(v_mark)
                    elif k == "Acc":
                        val = item.get("Acc", "")
                        if not val:
                            res = str(item.get("Result", "")).upper()
                            val = "OK" if "ACC" in res or "OK" in res else ""
                        row_vals.append(val)
                    elif k == "Rej":
                        val = item.get("Rej", "")
                        if not val:
                            res = str(item.get("Result", "")).upper()
                            val = "NG" if "REJ" in res or "NG" in res or "RE" in res else ""
                        row_vals.append(val)
                    elif k == "Deg":
                        row_vals.append(str(item.get("Deg", "")).strip())
                    else: 
                        val = str(item.get(k, "")).strip()
                        if (k in ["Dwg", "ISO"] and item.get('is_merged_iso')) or (k == "Joint" and item.get('is_merged_joint')): val = ""
                        row_vals.append(val)
            elif mode == "PT":
                for k in self.pt_column_keys:
                    if k == "selected": row_vals.append(v_mark)
                    else: 
                        val = str(item.get(k, "")).strip()
                        if (k in ["Dwg", "ISO"] and item.get('is_merged_iso')) or (k == "Joint" and item.get('is_merged_joint')): val = ""
                        row_vals.append(val)
            elif mode == "PAUT":
                # PAUT use paut_column_keys
                for k in self.paut_column_keys:
                    if k == "No": row_vals.append(len(idx_map))
                    else: 
                        val = str(item.get(k, "")).strip()
                        if (k in ["Dwg", "ISO"] and item.get('is_merged_iso')) or (k == "Joint" and item.get('is_merged_joint')): val = ""
                        row_vals.append(val)
            else: # PMI
                for k in self.column_keys:
                    if k == "_status": row_vals.append(st_mark)
                    elif k == "selected": row_vals.append(v_mark)
                    elif k in ["Ni", "Cr", "Mo", "Mn"]:
                        row_vals.append(f"{self.to_float(item.get(k, 0)):.2f}%")
                    else:
                        val = str(item.get(k, "")).strip()
                        if (k in ["Dwg", "ISO"] and item.get('is_merged_iso')) or (k == "Joint" and item.get('is_merged_joint')): val = ""
                        row_vals.append(val)

            row_tags = [str(idx), current_tag]
            if is_deficient:
                row_tags.append("deficient")
            tree.insert("", "end", values=tuple(row_vals), tags=tuple(row_tags))
            
        # [REACTIVE AUTO-RESET] Trigger if current view is clean but filters are active
        if mode == "PMI":
            has_pmi_filter = self.pmi_show_deficiency_only.get()
            if not has_pmi_filter and hasattr(self, 'element_filters'):
                for f in self.element_filters:
                    if f['min'].get().strip() or f['max'].get().strip():
                        has_pmi_filter = True; break
            
            if local_pmi_deficient_count == 0 and has_pmi_filter:
                self.reset_pmi_filters(refresh=True, clear_search=False)
                return

        if mode == "PMI" and hidden_by_def_count > 0:
            sel_str = ", ".join(list(set(selected_locs_names))) if selected_locs_names else "전체"
            self.log(f"ℹ️ '{sel_str}' 결과 중 {hidden_by_def_count}건이 '함량 미달만 보기' 옵션으로 인해 숨겨져 있습니다.")
            
        # [REMOVED] Legacy tab switching logic (Preview is now permanently visible in splitscreen)
        # if switch_tab:
        #     if mode == "RT":
        #         self.rt_tab_notebook.select(self.rt_tab_preview)
        #     elif mode == "PT":
        #         self.pt_tab_notebook.select(self.pt_tab_preview)
        #     elif mode == "PAUT":
        #         pass 
        #     else:
        #         self.tab_notebook.select(self.tab_preview)

    def _browse_dir(self, var):
        path = filedialog.askdirectory(initialdir=var.get() or RESOURCE_DIR)
        if path: var.set(path)

    def _update_file_info(self, mode, path):
        """Extracts and formats file metadata for the UI header."""
        if not path or not os.path.exists(path):
            self.file_info_vars[mode].set("📄 파일을 선택해주세요.")
            return

        try:
            fname = os.path.basename(path)
            fsize = os.path.getsize(path) / 1024 # KB
            mtime = os.path.getmtime(path)
            dt_str = datetime.datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M')
            
            size_str = f"{fsize:.1f}KB" if fsize < 1024 else f"{fsize/1024:.2f}MB"
            self.file_info_vars[mode].set(f"📄 파일: {fname}  |  📏 크기: {size_str}  |  📅 수정일: {dt_str}")
        except Exception as e:
            self.file_info_vars[mode].set(f"⚠️ 파일 정보 읽기 오류: {str(e)}")

    def _sync_all_file_infos(self):
        """Updates all mode headers from current config paths."""
        modes = [
            ('PMI', self.target_file_path.get()),
            ('RT', self.rt_target_file_path.get()),
            ('PT', self.pt_target_file_path.get()),
            ('PAUT', self.paut_target_file_path.get())
        ]
        for mode, path in modes:
            self._update_file_info(mode, path)

    def _browse_file(self, var, types):
        path = filedialog.askopenfilename(initialdir=os.path.dirname(var.get() or BASE_DIR), filetypes=types)
        if path: 
            var.set(path)
            # Find which mode this belongs to and update info
            for mode, m_var in [('PMI', self.target_file_path), ('RT', self.rt_target_file_path), 
                               ('PT', self.pt_target_file_path), ('PAUT', self.paut_target_file_path)]:
                if var == m_var:
                    self._update_file_info(mode, path)
                    break 

    def show_bulk_update_dialog(self, mode="PMI"):
        """선택된 항목들을 필터 조건에 따라 일괄 변경하는 다이얼로그 표시"""
        target_data = self.rt_extracted_data if mode == "RT" else (self.pt_extracted_data if mode == "PT" else self.extracted_data)
        
        if not target_data:
            messagebox.showwarning("알림", f"{mode} 데이터가 없습니다.")
            return

        # 체크(●)된 항목의 인덱스 추출
        selected_indices = [idx for idx, item in enumerate(target_data) if item.get('selected', True)]
        if not selected_indices:
            messagebox.showwarning("항목 미선택", "일괄 변경할 항목을 체크(선택)해주세요.")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title(f"{mode} 선택 항목 일괄 변경")
        dialog.geometry("450x570")
        dialog.configure(background="#f9fafb")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.geometry(f"+{self.root.winfo_x() + 100}+{self.root.winfo_y() + 50}")

        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill='both', expand=True)

        ttk.Label(main_frame, text="1. 대상 필터링 (선택 사항)", font=("Malgun Gothic", 10, "bold")).pack(anchor='w', pady=(0, 10))
        
        filter_frame = ttk.Frame(main_frame)
        filter_frame.pack(fill='x', pady=(0, 10))
        
        # 모드별 필터 가능 열 설정
        if mode == "RT":
            filter_cols = ["전체(필터 없음)", "Drawing No.", "Film Ident. No.", "Film Location", "Welder No", "Remarks"]
            key_map = {"Drawing No.": "Dwg", "Film Ident. No.": "Joint", "Film Location": "Loc", "Welder No": "Welder", "Remarks": "Remarks"}
        elif mode == "PT":
            filter_cols = ["전체(필터 없음)", "ISO Drawing No.", "Joint", "NPS", "Thk.", "Material", "Welder"]
            key_map = {"ISO Drawing No.": "Dwg", "Joint": "Joint", "NPS": "NPS", "Thk.": "Thk.", "Material": "Material", "Welder": "Welder"}
        else: # PMI
            filter_cols = ["전체(필터 없음)", "ISO/DWG", "Joint No", "Test Location", "Ni", "Cr", "Mo", "Grade"]
            key_map = {"ISO/DWG": "Dwg", "Joint No": "Joint", "Test Location": "Loc", "Ni": "Ni", "Cr": "Cr", "Mo": "Mo", "Grade": "Grade"}

        ttk.Label(filter_frame, text="필터 열:").grid(row=0, column=0, sticky='w')
        col_var = tk.StringVar(value="전체(필터 없음)")
        col_combo = ttk.Combobox(filter_frame, textvariable=col_var, state='readonly', width=20, values=filter_cols)
        col_combo.grid(row=0, column=1, padx=5, sticky='w')
        
        ttk.Label(filter_frame, text="필터 값:").grid(row=1, column=0, sticky='w', pady=5)
        val_var = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=val_var, width=23).grid(row=1, column=1, padx=5, sticky='w', pady=5)
        
        # [NEW] Filter Integration Checkboxes
        only_def_var = tk.BooleanVar(value=False)
        only_visible_var = tk.BooleanVar(value=True) # Default to True for convenience
        
        filter_options_frame = ttk.Frame(main_frame)
        filter_options_frame.pack(fill='x', pady=(0, 10))
        
        if mode == "PMI":
            tk.Checkbutton(filter_options_frame, text="⚠️ 함량 미달인 항목만 변경 적용", variable=only_def_var, 
                           background="#f9fafb", font=("Malgun Gothic", 9)).pack(anchor='w')
        
        tk.Checkbutton(filter_options_frame, text="🔍 현재 화면에 보이는(필터링된) 항목만 변경", variable=only_visible_var, 
                       background="#f9fafb", font=("Malgun Gothic", 9)).pack(anchor='w')
        
        ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=5)

        ttk.Label(main_frame, text="2. 변경할 값 입력 (입력된 항목만 반영)", font=("Malgun Gothic", 10, "bold")).pack(anchor='w', pady=(0, 10))
        
        fields_frame = ttk.Frame(main_frame)
        fields_frame.pack(fill='both', expand=True)
        
        # 모드별 입력 필드 구성
        if mode == "RT":
            labels = [("Dwg:", "Dwg"), ("Film Ident:", "Joint"), ("Loc:", "Loc"), ("Grade:", "Grade"), ("Welder:", "Welder"), ("Remarks:", "Remarks"), ("Date:", "Date")]
        elif mode == "PT":
            labels = [("Dwg:", "Dwg"), ("Joint:", "Joint"), ("NPS:", "NPS"), ("Thk:", "Thk."), ("Material:", "Material"), ("Welder:", "Welder"), ("Weld Type:", "WType"), ("Result:", "Result"), ("Date:", "Date")]
        else: # PMI
            labels = [("Ni (%):", "Ni"), ("Cr (%):", "Cr"), ("Mo (%):", "Mo"), ("Location:", "Loc"), ("Grade:", "Grade"), ("Date:", "Date")]

        entries = {}
        for i, (lbl, key) in enumerate(labels):
            ttk.Label(fields_frame, text=lbl).grid(row=i, column=0, sticky='e', pady=3, padx=5)
            entries[key] = ttk.Entry(fields_frame, width=20)
            entries[key].grid(row=i, column=1, sticky='w', pady=3)

        def apply_action():
            filter_col = col_var.get()
            filter_val = val_var.get().strip().lower()
            target_key = key_map.get(filter_col)
            
            update_spec = {}
            for key, entry in entries.items():
                val = entry.get().strip()
                if val:
                    if mode == "PMI" and key not in ["No", "Joint", "Loc", "Grade", "Date", "Dwg", "_status", "ST", "V", "selected", "order_index"]:
                        if val.startswith(('+', '-')):
                            try: update_spec[key] = ('relative', float(val))
                            except: update_spec[key] = ('absolute', self.to_float(val))
                        else:
                            update_spec[key] = ('absolute', self.to_float(val))
                    else:
                        update_spec[key] = ('absolute', val)
            
            if not update_spec:
                messagebox.showwarning("입력 부족", "변경할 값을 하나 이상 입력해주세요.")
                return

            count = 0
            for idx in selected_indices:
                item = target_data[idx]
                match = True
                if target_key:
                    curr_val = str(item.get(target_key, "")).strip().lower()
                    if filter_val not in curr_val: match = False
                
                # [NEW] Check for deficiency if requested
                if match and mode == "PMI" and only_def_var.get():
                    is_deficient = False
                    for f_item in self.element_filters:
                        f_key = f_item['key'].get().strip()
                        if not f_key: continue
                        f_min = round(self.to_float(f_item['min'].get()), 2)
                        f_max = round(self.to_float(f_item['max'].get()), 2)
                        val = round(self.to_float(self._get_val_ci(item, f_key)), 2)
                        if (f_min > 0 and val < f_min) or (f_max > 0 and val > f_max):
                            is_deficient = True; break
                    if not is_deficient: match = False
                
                # [NEW] Check for sidebar visibility (Date & Loc)
                if match and only_visible_var.get():
                    if not item.get('date_filtered', True):
                        match = False
                    
                    if match and mode == "PMI" and hasattr(self, 'pmi_loc_listbox'):
                        # Get active sidebar locs
                        active_locs = set()
                        for i in range(self.pmi_loc_listbox.size()):
                            v = self.pmi_loc_listbox.get(i)
                            if v.startswith("[v] "): active_locs.add(v.replace("[v] ", "").lower())
                        
                        if active_locs:
                            loc_val = str(item.get('Loc', '')).lower()
                            if loc_val not in active_locs:
                                match = False

                if match:
                    for k, spec in update_spec.items():
                        up_mode, up_val = spec
                        if up_mode == 'relative':
                            item[k] = self.to_float(item.get(k, 0)) + up_val
                        else:
                            item[k] = up_val
                    
                    # [NEW] Re-calculate Grade ONLY if NOT manually provided
                    if "Grade" not in update_spec:
                        new_grade = self.check_material_grade(item)
                        if new_grade:
                            item['Grade'] = new_grade
                    
                    count += 1
            
            self.populate_preview(target_data, switch_tab=False, mode=mode)
            messagebox.showinfo("일괄 변경 완료", f"총 {count}개의 항목이 업데이트되었습니다.")
            dialog.destroy()

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=(20, 0))
        ttk.Button(btn_frame, text="적용 (Apply)", command=apply_action).pack(side='right', padx=5)
        ttk.Button(btn_frame, text="취소 (Cancel)", command=dialog.destroy).pack(side='right')

    # --- Integrated Verification Logic ---

    def to_float(self, val):
        try:
            s = str(val).replace('%', '').strip()
            if not s or s.lower() == 'nan': return 0.0
            return float(s)
        except: return 0.0

    def _get_val_ci(self, item, key):
        """대소문자 구분 없이 딕셔너리에서 값 찾기 (Cr vs CR 등)"""
        if not key or not isinstance(item, dict): return None
        if key in item: return item[key]
        k_lower = key.lower()
        for k in item.keys():
            if str(k).lower() == k_lower:
                return item[k]
        return None

    def check_material_grade(self, row_data):
        """jjchRFIPMI.py에서 이식된 10% 여유치 판정 로직"""
        cr = self.to_float(self._get_val_ci(row_data, 'Cr'))
        ni = self.to_float(self._get_val_ci(row_data, 'Ni'))
        mo = self.to_float(self._get_val_ci(row_data, 'Mo'))
        mn = self.to_float(self._get_val_ci(row_data, 'Mn'))
        
        margin = 0.1 # 10% 여유
        
        # 1. SUS 316 (Cr:16~18 / Ni:10~14 / Mo:2~3)
        if (16.0*(1-margin) <= cr <= 18.0*(1+margin)) and (10.0*(1-margin) <= ni <= 14.0*(1+margin)) and (2.0*(1-margin) <= mo <= 3.0*(1+margin)):
            return "SS316"
        
        # 2. Duplex (Cr:22~23 / Mo:3~3.5 / Ni:4.5~6.5 / Mn:2.0이하)
        if (22.0*(1-margin) <= cr <= 23.0*(1+margin)) and (4.5*(1-margin) <= ni <= 6.5*(1+margin)) and (3.0*(1-margin) <= mo <= 3.5*(1+margin)) and (mn <= 2.2):
            return "DUPLEX"

        # 3. SUS 310 (Cr:24~26 / Ni:19~22)
        if (24.0*(1-margin) <= cr <= 26.0*(1+margin)) and (19.0*(1-margin) <= ni <= 22.0*(1+margin)):
            return "SS310"

        # 4. SUS 304 (Cr:18↑ / Ni:8↑ / Mo:0.5↓ / Mn:2.0↓)
        if (cr >= 16.2) and (ni >= 7.2) and (mo <= 0.55) and (mn <= 2.2):
            return "SS304"

        return None # 판정 불가 시 원본 사용

    def normalize_iso(self, val):
        """정렬 및 그룹화 시와 동일하게 정규화된 값으로 비교하여 도면번호 일관성 유지"""
        s_val = str(val).strip()
        try:
            if re.match(r'^\d+(\.\d+)?$', s_val):
                f_val = float(s_val)
                if f_val == int(f_val): return str(int(f_val))
                return str(f_val)
        except: pass
        return s_val.lower()

    # --- Excel Helper Logic ---

    def find_image_smart(self, keyword, exclude_keyword=None):
        def _search_in_folder(folder_path):
            if not folder_path or not os.path.exists(folder_path): return None
            candidates = glob.glob(os.path.join(folder_path, "*.*"))
            valid_extensions = ['.PNG', '.JPG', '.JPEG', '.BMP', '.GIF']
            for path in candidates:
                fname = os.path.basename(path).upper(); ext = os.path.splitext(path)[1].upper()
                if ext not in valid_extensions: continue 
                if keyword.upper() in fname:
                    if exclude_keyword and exclude_keyword.upper() in fname: continue
                    return path 
            return None

        # 1. UI에서 설정한 폴더에서 먼저 검색
        found = _search_in_folder(self.logo_folder_path.get())
        if found: return found
        
        # 2. PyInstaller 묶음(실행파일 내부 임시 폴더)에서 검색 (Standalone 지원)
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            found = _search_in_folder(sys._MEIPASS)
            if found: return found
            
        return None

    def place_image_freely(self, ws, img_path, anchor_cell_str, w, h, x_offset, y_offset):
        try:
            if not img_path or not os.path.exists(img_path): return
            # [NEW] Check for valid anchor cell string
            if not anchor_cell_str or not isinstance(anchor_cell_str, str) or len(anchor_cell_str) < 2:
                self.log(f"[WARNING] {os.path.basename(img_path)} 배치 실패: 잘못된 앵커 셀 ({anchor_cell_str})")
                return

            original = PILImage.open(img_path).convert("RGBA")
            # [NEW] Ensure w, h are positive
            w = max(1, int(w)); h = max(1, int(h))
            resized = original.resize((w, h), PILImage.Resampling.LANCZOS)
            temp_name = f"temp_{os.path.basename(img_path)}"
            temp_full_path = os.path.join(tempfile.gettempdir(), temp_name)
            resized.save(temp_full_path)
            
            img = XLImage(temp_full_path); img.width = w; img.height = h
            col_str, row_num = coordinate_from_string(anchor_cell_str)
            col_idx = max(0, column_index_from_string(col_str) - 1)
            row_idx = max(0, row_num - 1) 
            # [FIX] EMU offsets and dimensions MUST be non-negative
            emu_x = max(0, int(x_offset * 9525)); emu_y = max(0, int(y_offset * 9525))
            emu_w = max(1, int(w * 9525)); emu_h = max(1, int(h * 9525))
            
            marker = AnchorMarker(col=col_idx, colOff=emu_x, row=row_idx, rowOff=emu_y)
            size = XDRPositiveSize2D(cx=emu_w, cy=emu_h)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(img)
        except Exception as e: self.log(f"[WARNING] {os.path.basename(img_path)} 배치 실패: {e}")

    def add_logos_to_sheet(self, ws, is_cover=False, clear_existing=True):
        if clear_existing:
            try: ws._images = [] 
            except: pass
        ctx = "COVER" if is_cover else "DATA"
        def _get_params(key_prefix):
            return (self.config.get(f"{key_prefix}_{ctx}_PATH", ""), self.config[f"{key_prefix}_{ctx}_ANCHOR"], self.config[f"{key_prefix}_{ctx}_W"], self.config[f"{key_prefix}_{ctx}_H"], self.config[f"{key_prefix}_{ctx}_X"], self.config[f"{key_prefix}_{ctx}_Y"])
        
        # 1. SITCO
        path, anchor, w, h, x, y = _get_params("SITCO")
        if not path or not os.path.exists(path): path = self.find_image_smart("SITCO")
        self.place_image_freely(ws, path, anchor, w, h, x, y)
        
        # 2. SEOUL
        path, anchor, w, h, x, y = _get_params("SEOUL")
        if not path or not os.path.exists(path): path = self.find_image_smart("서울검사")
        self.place_image_freely(ws, path, anchor, w, h, x, y)
        
        # 3. FOOTER
        path, anchor, w, h, x, y = _get_params("FOOTER")
        if not path or not os.path.exists(path):
            path = self.find_image_smart("바닥글", exclude_keyword="PMI-1")
            if not path: path = self.find_image_smart("PMI", exclude_keyword="PMI-1")
        self.place_image_freely(ws, path, anchor, w, h, x, y)
        
        # 4. FOOTER_PT (Left)
        path, anchor, w, h, x, y = _get_params("FOOTER_PT")
        if not path or not os.path.exists(path):
            path = self.find_image_smart("PMI갑") if is_cover else None
            if not path: path = self.find_image_smart("PMI-1")
            if not path: path = self.find_image_smart("PT")
        self.place_image_freely(ws, path, anchor, w, h, x, y)

    def force_print_settings(self, ws, context="DATA"):
        try:
            # Context-specific logic for print area
            mode = "PMI" # Default
            # Attempt to detect mode from active tab or context
            try:
                active_tab = self.tab_control.index("current")
                if active_tab == 1: mode = "RT"
                elif active_tab == 2: mode = "PT"
                elif active_tab == 3: mode = "PAUT"
            except: pass

            if context == "COVER":
                # [NEW] Highly Dynamic print area for Gapji
                key = f"{mode}_GAPJI_PRINT_END_ROW" if mode != "PMI" else "GAPJI_PRINT_END_ROW"
                end_r = int(self.config.get(key, 51))
                if end_r > 0:
                    ws.print_area = f'A1:T{end_r}'
            else:
                # [NEW] Highly Dynamic print area for Eulji
                key = f"{mode}_PRINT_END_ROW" if mode != "PMI" else "PRINT_END_ROW"
                end_r = int(self.config.get(key, 47))
                if end_r > 0:
                    ws.print_area = f'A1:M{end_r}'
            ws.page_setup.paperSize = 9; ws.page_setup.orientation = 'portrait'
            
            # [REFINED] 사용자 설정 배율 및 여백 적용
            ws.page_setup.scale = int(float(self.config.get(f'PRINT_SCALE_{context}', 95)))
            ws.print_options.horizontalCentered = True; ws.print_options.verticalCentered = True
            
            ws.page_margins.top = float(self.config.get(f'MARGIN_{context}_TOP', 0.2))
            ws.page_margins.bottom = float(self.config.get(f'MARGIN_{context}_BOTTOM', 0.2))
            ws.page_margins.left = float(self.config.get(f'MARGIN_{context}_LEFT', 0.5))
            ws.page_margins.right = float(self.config.get(f'MARGIN_{context}_RIGHT', 0.3))
        except Exception as e:
            print(f"[WARNING] Print settings failed: {e}")

    def apply_custom_dimensions(self, ws, context):
        """[NEW] 사용자의 행/열 조절 설정을 파싱하여 적용"""
        try:
            # Row Adjustment
            row_range_str = self.config.get(f"CUSTOM_ROWS_{context}", "").strip()
            if row_range_str:
                height = float(self.config.get(f"CUSTOM_ROW_HEIGHT_{context}", 16.5))
                for part in row_range_str.split(','):
                    if '-' in part:
                        start, end = map(int, part.split('-'))
                        for r in range(start, end + 1):
                            ws.row_dimensions[r].height = height
                    else:
                        ws.row_dimensions[int(part)].height = height
            
            # Column Adjustment
            col_range_str = self.config.get(f"CUSTOM_COLS_{context}", "").strip()
            if col_range_str:
                width = float(self.config.get(f"CUSTOM_COL_WIDTH_{context}", 10.0))
                for part in col_range_str.split(','):
                    if '-' in part:
                        start_letter, end_letter = part.split('-')
                        for c in range(column_index_from_string(start_letter), column_index_from_string(end_letter) + 1):
                            col_letter = openpyxl.utils.get_column_letter(c)
                            ws.column_dimensions[col_letter].width = width
                    else:
                        ws.column_dimensions[part.strip().upper()].width = width
        except Exception as e:
            print(f"[DEBUG] Custom dimension failed: {e}")

    def safe_set_value(self, ws, coord, value, align=None):
        """병합된 셀이라도 기준 셀(Top-Left)을 찾아 안전하게 값을 입력합니다. (성능 최적화 버전)"""
        try:
            # 전달받은 인자가 이미 셀 객체인 경우
            if hasattr(coord, 'coordinate'):
                target_cell = coord
                coord_str = target_cell.coordinate
            else:
                target_cell = ws[coord]
                coord_str = coord

            # 해당 셀이 MergedCell(병합 영역의 기준 셀이 아닌 곳)인 경우에만 탐색 수행
            if isinstance(target_cell, MergedCell):
                for m_range in ws.merged_cells.ranges:
                    if coord_str in m_range:
                        target_cell = ws.cell(row=m_range.min_row, column=m_range.min_col)
                        break
            
            target_cell.value = value
            if align:
                target_cell.alignment = Alignment(horizontal=align, vertical='center')
        except Exception: pass

    def safe_merge_cells(self, ws, start_row, start_column, end_row, end_column):
        """이미 병합된 영역이 있는지 확인하고 안전하게 병합을 수행합니다."""
        try:
            # 병합하려는 영역이 이미 다른 병합 영역과 겹치는지 체크
            # openpyxl은 겹치는 병합을 허용하지 않으므로, 겹치면 해당 영역을 풀거나 스킵해야 함
            ws.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
        except Exception as e:
            # 겹치는 경우 기존 병합을 해제하고 시도하거나, 로그만 남김
            self.log(f"⚠️ 병합 실패 ({start_row},{start_column}~{end_row},{end_column}): {e}")

    def set_eulji_headers(self, ws):
        headers = ["NI", "CR", "MO"]
        data_font = Font(size=9); header_row = self.config['START_ROW']
        for i, val in enumerate(headers):
            col = 8 + i
            cell = ws.cell(row=header_row, column=col)
            self.safe_set_value(ws, cell.coordinate, val, align='center')
            cell.font = data_font
        
        materials = "SS304,SS304L,SS316,SS316L,SS321,SS347,SS410,SS430,DUPLEX,MONEL,INCONEL,ER308,ER308L,ER309,ER309L,ER316,ER316L,ER347,ER2209,WP316,WP316L,TP316,TP316L,F316L,A182-F316L,A312-TP316L"
        dv_q = DataValidation(type="list", formula1=f'"{materials}"', allow_blank=True)
        ws.add_data_validation(dv_q)
        for r in range(self.config['START_ROW'], self.config['DATA_END_ROW'] + 1):
            target_l = ws.cell(row=r, column=13); dv_q.add(target_l) # 12 -> 13 (M)
            target_l.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center'); target_l.font = Font(size=8.5)
    
        # [NEW] K15에 하이픈 추가
        self.safe_set_value(ws, 'K15', "-")
        ws['K15'].alignment = Alignment(horizontal='center', vertical='center')
        
    def prepare_next_sheet(self, wb, source_sheet_idx, page_num):
        source_sheet = wb.worksheets[source_sheet_idx]; new_sheet = wb.copy_worksheet(source_sheet) 
        base_title = source_sheet.title.split('_')[0]; new_sheet.title = f"{base_title[:20]}_{page_num:03d}"
        self.force_print_settings(new_sheet, context="DATA"); self.add_logos_to_sheet(new_sheet, is_cover=False)
        self.apply_custom_dimensions(new_sheet, "DATA")
        for col_letter, col_dim in source_sheet.column_dimensions.items(): new_sheet.column_dimensions[col_letter].width = col_dim.width
        data_font = Font(size=9); grade_font = Font(size=8.5)
        for r in range(self.config['START_ROW'], self.config['DATA_END_ROW'] + 1):
            rd = new_sheet.row_dimensions[r] # [REMOVED] Hardcoded 20.55 override
        
        for r in range(self.config['START_ROW'] + 1, self.config['DATA_END_ROW'] + 1):
            for c in range(1, 14):
                cell = new_sheet.cell(row=r, column=c)
                cell.font = grade_font if c == 13 else data_font
                self.safe_set_value(new_sheet, cell, None)
        merged_to_clear = [rng for rng in new_sheet.merged_cells.ranges if rng.min_row >= self.config['START_ROW'] and rng.max_row <= self.config['DATA_END_ROW']]
        for rng in merged_to_clear: new_sheet.unmerge_cells(str(rng))
        
        # [FIX] 갑지 데이터 수식으로 연결 (첫번째 시트 참조)
        try:
            ws0 = wb.worksheets[0]
            if len(wb.worksheets) > 1:
                # 엑셀 수식을 사용하여 갑지의 값이 바뀌면 자동으로 바뀌게 설정
                self.safe_set_value(new_sheet, 'K5', f"='{ws0.title}'!L5")
                self.safe_set_value(new_sheet, 'M5', f"='{ws0.title}'!N5")
                self.safe_set_value(new_sheet, 'M8', f"='{ws0.title}'!N8")
                # [FIX] K5:M10 범위 글씨체 바탕, 크기 9 적용
                for r_idx in range(5, 11):
                    for c_idx in range(11, 14): # K(11) ~ M(13)
                        cell = new_sheet.cell(row=r_idx, column=c_idx)
                        cell.font = Font(name='바탕', size=9, bold=False)
                        # K5, K8는 줄바꿈 적용, 그 외는 셀 크기에 맞춤 적용
                        if (r_idx == 5 or r_idx == 8) and c_idx == 11:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
        except: pass
        
        # [NEW] K15에 하이픈 추가
        self.safe_set_value(new_sheet, 'K15', "-")
        new_sheet['K15'].alignment = Alignment(horizontal='center', vertical='center')
        
        new_sheet.column_dimensions['M'].width = 18.0
        return new_sheet

    def extract_only(self, show_msg=True):
        """데이터만 추출하여 리스트와 미리보기에 반영 (PMI/RT 대응)"""
        # 현재 활성 탭에 따른 모드 결정
        try:
            tab_idx = self.mode_notebook.index("current") # Main notebook (PMI, RT, PT...)
            if tab_idx == 1: mode = "RT"
            elif tab_idx == 2: mode = "PT"
            else: mode = "PMI"
        except: mode = "PMI"

        target_file = self.pt_target_file_path.get() if mode == "PT" else (self.rt_target_file_path.get() if mode == "RT" else self.target_file_path.get())
        if not target_file:
            messagebox.showwarning("파일 미선택", f"{mode} 데이터 파일을 선택해주세요.")
            return False
            
        self.log(f"🔍 {mode} 데이터 추출 시작: {os.path.basename(target_file)}")
        
        # [NEW] Extract date from filename
        fname = os.path.basename(target_file)
        # More flexible regex for dates like 2024.03.05, 24.03.05, 24_03_05, 240305, etc.
        date_match = re.search(r'(\d{4}[-._]\d{2}[-._]\d{2}|\d{2}[-._]\d{2}[-._]\d{2}|\d{8}|\d{6})', fname)
        extracted_date = date_match.group(0) if date_match else ""
        
        # Standardize formatting
        if extracted_date:
            # Remove separators like . or _
            clean_date = re.sub(r'[-._]', '', extracted_date)
            if len(clean_date) == 8:
                extracted_date = f"{clean_date[:4]}-{clean_date[4:6]}-{clean_date[6:]}"
            elif len(clean_date) == 6:
                extracted_date = f"20{clean_date[:2]}-{clean_date[2:4]}-{clean_date[4:]}"
        
        if not extracted_date:
            extracted_date = datetime.datetime.now().strftime("%Y-%m-%d")
            self.log(f"⚠️ 파일명에서 날짜를 찾지 못해 오늘 날짜({extracted_date})로 설정합니다.")
        else:
            self.log(f"📅 파일 날짜 인식 성공: {extracted_date}")

        self.progress['value'] = 0
        all_extracted_data = []

        # [AUTO COLUMN ADDITION] 
        if mode == "PMI" and hasattr(self, 'element_filters') and self.element_filters:
            new_keys = []
            for f_item in self.element_filters:
                f_k = f_item['key'].get().strip()
                if f_k and f_k.upper() not in ["CR", "NI", "MO", "MN"] and f_k.capitalize() not in self.column_keys and f_k.upper() not in self.column_keys and f_k not in self.column_keys:
                    new_keys.append(f_k)
            
            if new_keys:
                self.column_keys.extend(new_keys)
                current_cols = list(self.preview_tree["columns"])
                current_display = list(self.preview_tree.cget("displaycolumns"))
                if current_display[0] == "#all": current_display = list(current_cols)

                # Save existing configurations to survive the tree["columns"] reset
                configs = {}
                for c in current_cols:
                    configs[c] = {
                        "text": self.preview_tree.heading(c).get("text", c),
                        "width": self.preview_tree.column(c).get("width", 80)
                    }

                for k in new_keys:
                    if k not in current_cols: current_cols.append(k)
                    if k not in current_display: current_display.append(k)
                
                self.preview_tree["columns"] = tuple(current_cols)
                self.preview_tree["displaycolumns"] = tuple(current_display)
                
                # Re-apply all configs + new keys setup
                for c in current_cols:
                    header_txt = configs.get(c, {}).get("text", c)
                    w = configs.get(c, {}).get("width", 80)
                    self.preview_tree.heading(c, text=header_txt, anchor='center', command=lambda _c=c: self.sort_by_column(_c, mode="PMI"))
                    self.preview_tree.column(c, width=w, anchor='center', stretch=False)
        
        try:
            target_input = self.sequence_filter.get().strip()
            target_no_list = [x.strip() for x in target_input.replace(',', ' ').split() if target_input and x.strip()] if target_input else []
            
            with pd.ExcelFile(target_file) as xls:
                for s_idx, sheet_name in enumerate(xls.sheet_names):
                    self.log(f"📄 시트 스캔: {sheet_name}")
                    try: temp_df = pd.read_excel(xls, sheet_name=sheet_name, header=None, nrows=50)
                    except: continue
                    header_idx = None
                    for i, row in temp_df.iterrows():
                        row_str = str(row.values).upper()
                        if mode == "RT":
                            if "FILM" in row_str or "DEFECT" in row_str or "IQI" in row_str:
                                header_idx = i; break
                        elif mode == "PT":
                            if (("LINE" in row_str or "ISO" in row_str or "DWG" in row_str) and ("JOINT" in row_str or "WELD" in row_str)):
                                header_idx = i; break
                        else:
                            if ("CR" in row_str and "NI" in row_str) or ("CHROMIUM" in row_str):
                                header_idx = i; break
                    if header_idx is None: continue
                    
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=header_idx)
                    
                    # [NEW] [Refinement] Two-line header detection and merging
                    if not df.empty and len(df) > 0:
                        row0_str = " ".join([str(x).upper() for x in df.iloc[0] if pd.notna(x)])
                        if any(k in row0_str for k in ["ORIGIN", "FACTOR", "WELDER1"]):
                            self.log(f"   ℹ️ [자동보정] 두 줄 헤더 감지됨 (시트: {sheet_name})")
                            new_cols = []
                            for col, val in zip(df.columns, df.iloc[0]):
                                c_txt = str(col) if pd.notna(col) and "Unnamed" not in str(col) else ""
                                v_txt = str(val) if pd.notna(val) else ""
                                new_cols.append(f"{c_txt} {v_txt}".strip())
                            df.columns = new_cols
                            df = df.iloc[1:].reset_index(drop=True)

                def _find_col(df, keywords, exclude=None):
                    # 1. 우선 정확히 일치하는 컬럼을 찾음
                    for col in df.columns:
                        c_up = str(col).upper().strip()
                        if exclude and any(ex in c_up for ex in exclude): continue
                        if any(k == c_up for k in keywords): return col
                    # 2. 포함된 컬럼을 찾되, NI의 경우 UNIT 등 오진 가능성 차단
                    for col in df.columns:
                        c_up = str(col).upper().strip()
                        if exclude and any(ex in c_up for ex in exclude): continue
                        if any(k in c_up for k in keywords):
                            if "NI" in keywords and ("UNIT" in c_up or "LINE" in c_up): continue
                            return col
                    return None
                if mode == "RT":
                    col_no = _find_col(df, ["NO.", "NO", "SEQ", "ITEM"])
                    col_dwg = _find_col(df, ["ISO", "DWG", "DRAWING", "LINE"])
                    col_joint = _find_col(df, ["JOINT", "WELD NO", "J/N", "FILM IDENT"])
                    col_loc = _find_col(df, ["LOCATION", "POSITION", "FILM LOC"])
                    col_grade = _find_col(df, ["GRADE", "DEG"]) # DEG for RT
                    col_welder = _find_col(df, ["WELDER", "W/N"])
                    col_remarks = _find_col(df, ["REMARKS", "REMARK", "비고"])
                    # RT Specifics
                    col_date = _find_col(df, ["DATE", "검사일"])
                    col_t = _find_col(df, ["T", "THICK", "THK"])
                    col_mat = _find_col(df, ["MAT", "MATERIAL"])
                    col_weld = _find_col(df, ["WELD", "TYPE"])
                    col_iqi = _find_col(df, ["IQI"])
                    col_sens = _find_col(df, ["SENS", "SENSITIVITY"])
                    col_den = _find_col(df, ["DEN", "DENSITY"])
                    col_acc = _find_col(df, ["ACC", "합격"])
                    col_rej = _find_col(df, ["REJ", "불합격"])
                    col_deg = _find_col(df, ["DEG", "물성", "수정브랜드", "GRADE"])
                    col_result = _find_col(df, ["RESULT", "판정"])
                    
                    defect_cols = {}
                    for i in range(1, 16):
                        # Support for D1, DEFECT1, and circled numbers ① to ⑮
                        circle_num = chr(9311 + i) # ① is 9312
                        c = _find_col(df, [f"D{i}", f"DEFECT{i}", circle_num, f"{i}"])
                        if c: defect_cols[f"D{i}"] = c
                elif mode == "PT":
                    col_no = _find_col(df, ["NO.", "NO", "SEQ", "ITEM"])
                    col_dwg = _find_col(df, ["ISO", "LINE", "DWG", "DRAWING"], exclude=["JOINT", "WELD"]) 
                    col_joint = _find_col(df, ["JOINT NO", "JOINT NUMBER"], exclude=["ISO", "LINE", "ITEM"])
                    if not col_joint:
                        col_joint = _find_col(df, ["JOINT", "WELD"], exclude=["ISO", "LINE", "ITEM", "WPS", "REPORT", "DATE", "TYPE"])
                    col_mat = _find_col(df, ["MAT", "MATERIAL", "재질", "CLASS", "M'TL"])
                    col_size = _find_col(df, ["SIZE", "NPS", "DIA", "INCH", "관경", "ORIGIN", "DI ORIGIN"])
                    col_thk = _find_col(df, ["THK", "THICK", "SCH", "두께"])
                    col_welder = _find_col(df, ["WELDER", "ID", "용접사", "WELDER1"])
                    col_wtype = _find_col(df, ["WELD TYPE", "WELDTYPE", "W.TYPE", "TYPE"], exclude=["JOINT"])
                    col_result = _find_col(df, ["RESULT", "결과", "판정"])
                    # Placeholder for Loc to maintain structure
                    col_loc = None 
                else:
                    col_cr = _find_col(df, ["CR", "CHROMIUM"]); col_ni = _find_col(df, ["NI", "NICKEL"])
                    col_mo = _find_col(df, ["MO", "MOLYBDENUM"]); col_mn = _find_col(df, ["MN", "MANGANESE"])
                    col_no = _find_col(df, ["NO.", "NO", "SEQ", "NUM", "POS", "ITEM"])
                    col_joint = _find_col(df, ["JOINT", "J/N", "JOINT NO", "PUNCH", "WELD NO"])
                    col_loc = _find_col(df, ["LOCATION", "TEST POSITION", "POINT", "AREA", "POSITION"])
                    col_dwg = _find_col(df, ["ISO", "DWG", "DRAWING", "LINE"])
                    col_grade_orig = _find_col(df, ["GRADE", "MATERIAL", "SPEC", "TYPE"])
                    
                    # [DYNAMIC] Scan for dynamically added filter elements (e.g., C, P, S)
                    custom_cols = {}
                    if hasattr(self, 'element_filters') and self.element_filters:
                        for f_item in self.element_filters:
                            f_key = f_item['key'].get().strip()
                            if f_key and f_key.upper() not in ["CR", "NI", "MO", "MN"]:
                                c_found = _find_col(df, [f_key.upper()])
                                if c_found is not None:
                                    custom_cols[f_key] = c_found

                # [NEW] Forward-fill (Fill Down) for ISO and Joint
                last_dwg = ""
                last_joint = ""

                for _, row in df.iterrows():
                    v_raw_no = str(row[col_no]).strip() if col_no is not None else str(_+1)
                    if target_no_list and v_raw_no not in target_no_list: continue
                    
                    # [NEW] Common Fill-down Logic
                    curr_dwg = str(row[col_dwg]).strip() if col_dwg is not None else ""
                    if (not curr_dwg or curr_dwg == "nan") and last_dwg: curr_dwg = last_dwg
                    if curr_dwg and curr_dwg != "nan": last_dwg = curr_dwg

                    curr_joint = str(row[col_joint]).strip() if col_joint is not None else ""
                    if (not curr_joint or curr_joint == "nan") and last_joint: curr_joint = last_joint
                    if curr_joint and curr_joint != "nan": last_joint = curr_joint
                    elif not curr_joint or curr_joint == "nan": curr_joint = v_raw_no

                    if mode == "RT":
                        # RT-specific Row Data
                        item_data = {
                            'No': v_raw_no, 
                            'Date': str(row[col_date]).strip() if col_date is not None else (extracted_date if extracted_date else ""),
                            'Dwg': curr_dwg, 'Joint': curr_joint,
                            'Loc': str(row[col_loc]).strip() if col_loc is not None else "",
                            'Acc': str(row[col_acc]).strip() if col_acc is not None else "",
                            'Rej': str(row[col_rej]).strip() if col_rej is not None else "",
                            'Deg': str(row[col_deg]).strip() if col_deg is not None else "",
                            'Welder': str(row[col_welder]).strip() if col_welder is not None else "",
                            'Remarks': str(row[col_remarks]).strip() if col_remarks is not None else "",
                            'T': str(row[col_t]).strip() if col_t is not None else "",
                            'Mat': str(row[col_mat]).strip() if col_mat is not None else "",
                            'Weld': str(row[col_weld]).strip() if col_weld is not None else "",
                            'IQI': str(row[col_iqi]).strip() if col_iqi is not None else "",
                            'Sens': str(row[col_sens]).strip() if col_sens is not None else "",
                            'Den': str(row[col_den]).strip() if col_den is not None else "",
                            'Result': str(row[col_result]).strip() if col_result is not None else "ACC",
                            'selected': True,
                            'order_index': len(self.rt_extracted_data) + len(all_extracted_data)
                        }
                        # Defects D1-D15 (Normalization)
                        for i in range(1, 16):
                            key = f"D{i}"
                            c = defect_cols.get(key)
                            val = str(row[c]).strip() if c is not None else ""
                            if val and val.lower() in ["v", "x", "o", "1", "√", "v"]: val = "√"
                            else: val = ""
                            item_data[key] = val
                            
                        # [NEW] 기존에 수동으로 추가된 컬럼이 있다면 해당 키도 포함하여 초기화
                        for k in self.rt_column_keys:
                            if k not in item_data and k != "selected":
                                item_data[k] = ""

                        all_extracted_data.append(item_data)
                    elif mode == "PT":
                        # PT-specific Row Data (Filter for ACC only)
                        res_str = str(row[col_result]).upper() if col_result is not None else "ACC"
                        is_pass = any(k in res_str for k in ["ACC", "OK", "ACCEPT", "합격"])
                        is_fail = any(k in res_str for k in ["REJ", "RW", "FAIL", "UNACC"])
                        if is_pass and not is_fail:
                            size_v = str(row[col_size]).strip() if col_size is not None else ""
                            thk_v = str(row[col_thk]).strip() if col_thk is not None else ""
                            # SCH -> Thk 변환
                            thk_converted = convert_sch_to_thk(size_v, thk_v)
                            
                            item_row = {
                                'No': v_raw_no, 'Date': extracted_date, 'Dwg': curr_dwg, 'Joint': self.force_two_digit(curr_joint),
                                'NPS': size_v, 'Thk.': thk_converted, 
                                'Material': self.fix_material_name(row[col_mat]) if col_mat is not None else "",
                                'Welder': str(row[col_welder]).strip() if col_welder is not None else "",
                                'WType': str(row[col_wtype]).strip() if col_wtype is not None else "",
                                'Result': "Acc",
                                'selected': True,
                                'order_index': len(self.pt_extracted_data) + len(all_extracted_data)
                            }
                            # [NEW] 기존 수동 추가 컬럼 보존
                            for k in self.pt_column_keys:
                                if k not in item_row and k != "selected":
                                    item_row[k] = ""
                            all_extracted_data.append(item_row)
                    else:
                        # PMI-specific Row Data
                        v_cr = self.to_float(row[col_cr])
                        if v_cr > 0 or (v_raw_no != "" and v_raw_no != "nan"):
                            v_ni = self.to_float(row[col_ni])
                            v_mo = self.to_float(row[col_mo]) if col_mo is not None else 0.0
                            v_mn = self.to_float(row[col_mn]) if col_mn is not None else 0.0
                            orig_grade = str(row[col_grade_orig]).strip() if col_grade_orig is not None else ""
                            
                            final_grade = orig_grade
                            if self.auto_verify.get():
                                detected = self.check_material_grade({'Cr': v_cr, 'Ni': v_ni, 'Mo': v_mo, 'Mn': v_mn})
                                if detected: final_grade = detected
                            if not final_grade or final_grade == "nan":
                                final_grade = "SS316" if v_mo >= 1.5 else "SS304"

                            pmi_item = {
                                'No': v_raw_no, 
                                'Joint': curr_joint,
                                'Loc': str(row[col_loc]).strip() if col_loc is not None else "",
                                'Cr': v_cr, 'Ni': v_ni, 'Mo': v_mo, 'Mn': v_mn,
                                'Grade': final_grade, 'Dwg': curr_dwg,
                                'Date': extracted_date,
                                'selected': True,
                                'order_index': len(self.extracted_data) + len(all_extracted_data)
                            }
                            # Extract dynamically found elements
                            for c_key, c_idx in custom_cols.items():
                                pmi_item[c_key] = self.to_float(row[c_idx])
                            # [NEW] 기존 수동 추가 컬럼 보존
                            for k in self.column_keys:
                                if k not in pmi_item and k != "selected":
                                    pmi_item[k] = ""
                            all_extracted_data.append(pmi_item)
                self.progress['value'] = ((s_idx + 1) / len(xls.sheet_names)) * 50

            if not all_extracted_data:
                messagebox.showerror("오류", "추출된 데이터가 없습니다.")
                return False

            # Extraction Mode Filter (PMI Only)
            if mode == "PMI":
                ext_mode = self.extraction_mode.get()
                if ext_mode != "전체":
                    original_count = len(all_extracted_data)
                    if ext_mode == "SS304 만": all_extracted_data = [d for d in all_extracted_data if d['Grade'] == "SS304"]
                    elif ext_mode == "SS316 만": all_extracted_data = [d for d in all_extracted_data if d['Grade'] == "SS316"]
                    elif ext_mode == "DUPLEX 만": all_extracted_data = [d for d in all_extracted_data if d['Grade'] == "DUPLEX"]
                    elif ext_mode == "SS310 만": all_extracted_data = [d for d in all_extracted_data if d['Grade'] == "SS310"]
                    elif ext_mode == "미분류(기타) 만":
                        known_grades = ["SS304", "SS316", "DUPLEX", "SS310"]
                        all_extracted_data = [d for d in all_extracted_data if d['Grade'] not in known_grades]
                    
                    self.log(f"🔍 필터링 ({ext_mode}): {original_count}개 -> {len(all_extracted_data)}개")
                    if not all_extracted_data:
                        messagebox.showinfo("알림", f"'{ext_mode}'에 해당하는 데이터가 없습니다.")
                        return False

            # [NEW] 원소 함량 필터링 적용 (PMI 전용)
            if mode == "PMI" and hasattr(self, 'element_filters') and self.element_filters:
                for f_item in self.element_filters:
                    f_key = f_item['key'].get().strip()
                    if not f_key: continue
                    f_min = self.to_float(f_item['min'].get())
                    f_max = self.to_float(f_item['max'].get())
                    
                    if f_min > 0 or f_max > 0:
                        original_count = len(all_extracted_data)
                        if f_min > 0 and f_max > 0:
                            all_extracted_data = [d for d in all_extracted_data if f_min <= d.get(f_key, d.get(f_key.capitalize(), 0.0)) <= f_max]
                        elif f_min > 0:
                            all_extracted_data = [d for d in all_extracted_data if d.get(f_key, d.get(f_key.capitalize(), 0.0)) >= f_min]
                        elif f_max > 0:
                            all_extracted_data = [d for d in all_extracted_data if d.get(f_key, d.get(f_key.capitalize(), 0.0)) <= f_max]
                        
                        if len(all_extracted_data) != original_count:
                             self.log(f"🔍 원소 필터 적용 ({f_key}: {f_min}~{f_max}): {original_count} -> {len(all_extracted_data)}건 남음")

            if mode == "PT":
                # [NEW] [Refinement] Duplicate removal for PT (ISO + Joint)
                seen = set()
                unique_data = []
                for d in all_extracted_data:
                    key = (str(d['Dwg']).strip(), str(d['Joint']).strip())
                    if key not in seen:
                        seen.add(key)
                        unique_data.append(d)
                all_extracted_data = unique_data

            if target_no_list:
                all_extracted_data.sort(key=lambda x: target_no_list.index(str(x['No'])) if str(x['No']) in target_no_list else 999999)

            # [CHANGE] Overwrite -> Accumulate + Auto Sort
            if mode == "RT":
                self.rt_extracted_data.extend(all_extracted_data)
                self.update_date_listbox("RT")
                self.sort_by_column("ISO Drawing No.", mode="RT") # Auto sort for RT
                total_count = len(self.rt_extracted_data)
            elif mode == "PT":
                self.pt_extracted_data.extend(all_extracted_data)
                self.update_date_listbox("PT")
                self.sort_by_column("ISO Drawing No.", mode="PT") # Auto sort for PT
                total_count = len(self.pt_extracted_data)
            else:
                self.extracted_data.extend(all_extracted_data)
                self.update_date_listbox("PMI")
                self.sort_by_column("ISO/DWG", mode="PMI") # Auto sort for PMI
                total_count = len(self.extracted_data)
            
            self.progress['value'] = 100
            if show_msg:
                self.log(f"✅ {mode} 데이터 누적 완료 (현재 총 {total_count} 건)")
                messagebox.showinfo("완료", f"데이터가 추가되었습니다.\n현재 목록에는 총 {total_count}건의 데이터가 있습니다.")
            return True
        except Exception as e:
            self.log(f"❌ {mode} 추출 오류: {e}")
            traceback.print_exc()
            return False

    def duplicate_selected_rows(self):
        """선택된 행들을 복제하여 데이터 리스트에 삽입"""
        selected = self.preview_tree.selection()
        if not selected: return
        
        # 시각적 순서대로 정렬 (위에서부터 복제)
        selected = sorted(list(selected), key=lambda x: self.preview_tree.index(x))
        
        new_items = []
        insert_pos = 0
        for item_id in selected:
            view_idx = self.preview_tree.index(item_id)
            actual_idx = self.item_idx_map[view_idx]
            new_item = self.extracted_data[actual_idx].copy()
            # No.는 중복 방지를 위해 비우거나 새로 부여 (여기서는 유지 후 나중에 rename 가능)
            new_items.append(new_item)
            insert_pos = actual_idx
            
        for item in reversed(new_items):
            self.extracted_data.insert(insert_pos + 1, item)
             
        self.populate_preview(self.extracted_data, switch_tab=False)
        self.log(f"👯 {len(new_items)}개 행 복제 완료")

    def save_column_widths(self):
        """현재 트리뷰의 컬럼 너비를 설정 파일에 저장"""
        widths = {}
        for col in self.preview_tree["columns"]:
            widths[col] = self.preview_tree.column(col, "width")
        
        self.config["PMI_COL_WIDTHS"] = widths
        self.save_settings()
        self.log("📏 컬럼 너비 설정 저장 완료")

    def apply_date_filter(self):
        """날짜 리스트박스 선택 상태를 데이터 모델에 반영"""
        selected_dates = []
        for i in range(self.date_listbox.size()):
            val = self.date_listbox.get(i)
            if val.startswith("[v]"):
                selected_dates.append(val.replace("[v] ", "").strip())
        
        if not selected_dates:
            messagebox.showwarning("필터 오류", "최소 하나 이상의 날짜를 선택해주세요.")
            return

        for item in self.extracted_data:
            item['date_filtered'] = (item.get('Date', '') in selected_dates)
        
        self.populate_preview(self.extracted_data, switch_tab=False)
        self.log(f"📅 날짜 필터 적용 완료: {len(selected_dates)}개 날짜 선택됨")

    def run_process(self):
        # [NEW] Ensure config values are correct types for comparison (prevent str vs int errors)
        for k in list(self.config.keys()):
            if k.endswith(('_ROW', '_IDX', '_SIZE')) or any(x in k for x in ['START', 'END', 'PAGE']):
                try: self.config[k] = int(float(self.config[k]))
                except: pass
            elif any(x in k for x in ['MARGIN', 'SCALE', 'RATIO', 'POS']):
                try: self.config[k] = float(self.config[k])
                except: pass

        # 현재 활성 탭에 따른 모드 결정
        try:
            tab_idx = self.mode_notebook.index("current")
            if tab_idx == 1: mode = "RT"
            elif tab_idx == 2: mode = "PT"
            else: mode = "PMI"
        except: mode = "PMI"

        if mode == "RT":
            target_file = self.rt_target_file_path.get()
            template_path = self.rt_template_file_path.get()
            data = self.rt_extracted_data
        elif mode == "PT":
            target_file = self.pt_target_file_path.get()
            template_path = self.pt_template_file_path.get()
            data = self.pt_extracted_data
        else:
            target_file = self.target_file_path.get()
            template_path = self.template_file_path.get()
            data = self.extracted_data
        
        if not template_path:
            messagebox.showwarning("파일 미선택", f"{mode} 양식(Template) 파일을 선택해주세요.")
            return
            
        if not target_file and not data:
            messagebox.showwarning("파일 미선택", f"{mode} 데이터 파일(Excel)을 선택하거나, 저장된 데이터를 불러와주세요.")
            return

        # [NEW] 템플릿 파일 존재 여부 및 기본 구조 확인
        if not os.path.exists(template_path):
            messagebox.showerror("오류", f"템플릿 파일을 찾을 수 없습니다:\n{template_path}")
            return
            
        self.save_settings()
        
        # 데이터가 비어있거나 새로 추출이 필요한 경우 수행
        if not data:
            if not self.extract_only(show_msg=False): return
            # Re-fetch data after extraction
            data = self.pt_extracted_data if mode == "PT" else (self.rt_extracted_data if mode == "RT" else self.extracted_data)
            
        # [NEW] 체크된 항목만 필터링 (기본값은 True)
        final_list = [d for d in data if d.get('selected', True) and d.get('date_filtered', True)]
        if not final_list:
            messagebox.showwarning("항목 미선택", f"선택된 {mode} 데이터가 없습니다. 미리보기에서 항목을 체크해주세요.")
            return

        if mode == "RT":
            self._run_rt_process(final_list, template_path)
        elif mode == "PT":
            self._run_pt_process(final_list, template_path)
        else:
            self._run_pmi_process(final_list, template_path)

    def _run_pmi_process(self, final_list, template_path):
        self.log(f"🚀 PMI 성적서 생성 시작 (총 {len(final_list)} 건)...")
        self.progress['value'] = 0
        all_extracted_data = final_list
        
        # [FIX] data_end_row 전역/설정으로 보강
        data_start_row = int(self.config.get('START_ROW', 17))
        data_end_row = int(self.config.get('DATA_END_ROW', 45))
        
        try:
            wb = openpyxl.load_workbook(template_path, keep_vba=True)
            if len(wb.worksheets) < 1:
                raise ValueError("선택한 템플릿 파일에 시트가 존재하지 않습니다.")

            if len(wb.worksheets) >= 1:
                ws0 = wb.worksheets[0]; self.add_logos_to_sheet(ws0, is_cover=True, clear_existing=False)
                self.force_print_settings(ws0, context="COVER") # [NEW] 갑지 전용 여백 적용
                
                # [NEW] 갑지 전용 여백 적용
                self.force_print_settings(ws0, context="COVER")
                # [NEW] Dynamic Gapji Border Range (Using Standardized Gapji Data Start/End)
                b_start = int(self.config.get('GAPJI_START_ROW', 23))
                b_end = int(self.config.get('GAPJI_DATA_END_ROW', 38))
                
                if b_start > 0 and b_end >= b_start:
                    for r in range(b_start, b_end + 1):
                        try:
                            cell_a = ws0.cell(row=r, column=1); eb = cell_a.border
                            cell_a.border = Border(left=medium_side, right=eb.right, top=eb.top, bottom=eb.bottom)
                        except: pass
                
                ws0['I35'].border = Border() # [FIX] I35 셀 선 제거
                self.safe_set_value(ws0, 'I35', None) 
                self.apply_custom_dimensions(ws0, "COVER") # [MOVED] Ensure it has the final say
            
            data_sheet_id = 1 if len(wb.worksheets) >= 2 else 0
            ws = wb.worksheets[data_sheet_id]; ws.title = f"{ws.title[:20]}_001"
            # 을지 기본 설정
            self.add_logos_to_sheet(ws, is_cover=False); self.force_print_settings(ws, context="DATA"); self.set_eulji_headers(ws)
            # self.apply_custom_dimensions(ws, "DATA") # [MOVED] To the end of process
            
            try:
                if len(wb.worksheets) >= 2:
                    # [FIX] 수식으로 연결 (갑지 L5 -> 을지 K5, 갑지 N5 -> 을지 M5, 갑지 N8 -> 을지 M8)
                    self.safe_set_value(ws, 'K5', f"='{ws0.title}'!L5")
                    self.safe_set_value(ws, 'M5', f"='{ws0.title}'!N5")
                    self.safe_set_value(ws, 'M8', f"='{ws0.title}'!N8")
                    # [FIX] K5:M10 범위 스타일 설정 (바탕, 크기 9)
                    for r_idx in range(5, 11):
                        for c_idx in range(11, 14):
                            cell = ws.cell(row=r_idx, column=c_idx)
                            cell.font = Font(name='바탕', size=9, bold=False)
                            if (r_idx == 5 or r_idx == 8) and c_idx == 11: # K5, K8 줄바꿈
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            else:
                                cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                    ws.column_dimensions['M'].width = 18.0
            except Exception as e:
                self.log(f"갑지 데이터 복사 실패: {e}")
            
            materials = "SS304,SS304L,SS316,SS316L,SS321,SS347,SS410,SS430,DUPLEX,MONEL,INCONEL,ER308,ER308L,ER309,ER309L,ER316,ER316L,ER347,ER2209,WP316,WP316L,TP316,TP316L,F316L,A182-F316L,A312-TP316L"
            dv_q = DataValidation(type="list", formula1=f'"{materials}"', allow_blank=True)
            # [NEW] 데이터 기입 전 기존 병합 영역 정리 (A-E 등 중복 병합 방지)
            def clear_merges_in_range(sheet, start_row, end_row):
                merged_ranges = list(sheet.merged_cells.ranges)
                for r in merged_ranges:
                    # 영역이 겹치면 해제 (aggressive clearing)
                    if r.min_row <= end_row and r.max_row >= start_row:
                        try: sheet.unmerge_cells(str(r))
                        except: pass

            ws = wb.worksheets[data_sheet_id]
            clear_merges_in_range(ws, self.config['START_ROW'], self.config['DATA_END_ROW'] + 20)
            ws.add_data_validation(dv_q)

            # [DYNAMIC ELEMENTS] Identify newly appended chemical elements
            custom_elements = [k for k in self.column_keys if k not in ["_status", "selected", "No", "Date", "Dwg", "Joint", "Loc", "Grade", "order_index", "V", "ST", "Ni", "Cr", "Mo"]]
            
            # [NEW] Dynamic Column Mapping for Elements
            ni_col = self.col_to_num(self.config.get('PMI_COL_NI', '8'))
            cr_col = self.col_to_num(self.config.get('PMI_COL_CR', '9'))
            mo_col = self.col_to_num(self.config.get('PMI_COL_MO', '10'))
            
            print_elements = [('Ni', ni_col), ('Cr', cr_col), ('Mo', mo_col)]
            if len(custom_elements) > 0: print_elements.append((custom_elements[0], 11))
            if len(custom_elements) > 1: print_elements.append((custom_elements[1], 12))

            current_row = self.config['START_ROW']; current_page = 1; data_ptr = 0
            while data_ptr < len(all_extracted_data):
                # 가용 행 수 계산 (DATA_END_ROW까지만 채움)
                rows_left = self.config['DATA_END_ROW'] - current_row + 1
                
                # 만약 공간이 전혀 없으면 새 시트로 전환
                if rows_left <= 0:
                    current_page += 1; ws = self.prepare_next_sheet(wb, data_sheet_id, current_page)
                    clear_merges_in_range(ws, self.config['START_ROW'], self.config['DATA_END_ROW'] + 20)
                    current_row = self.config['START_ROW']; ws.add_data_validation(dv_q)
                    rows_left = self.config['DATA_END_ROW'] - current_row + 1

                # [NEW] Dynamically print headers for the new elements precisely on the first row of each sheet
                if current_row == self.config['START_ROW'] and len(print_elements) > 3:
                    h_row = self.config['START_ROW'] - 1
                    for r_i in range(max(1, self.config['START_ROW'] - 5), self.config['START_ROW']):
                        if str(ws.cell(row=r_i, column=8).value).strip().upper() == "NI":
                            h_row = r_i; break
                    for val_k, c_idx in print_elements[3:]:
                        hdr_cell = ws.cell(row=h_row, column=c_idx)
                        self.safe_set_value(ws, hdr_cell.coordinate, val_k, align='center')
                        hdr_cell.font = Font(name=hdr_cell.font.name if hasattr(hdr_cell.font, 'name') else 'Arial', size=hdr_cell.font.size if hasattr(hdr_cell.font, 'size') else 9, bold=True)

                # [NEW DYNAMIC ALGORITHM] Determine block size based on the UI's manual merge logic
                # We start with the item at data_ptr. We scan forward to see how many subsequent items
                # legitimately belong to this same visual block.
                logical_block_size = 1
                base_item = all_extracted_data[data_ptr]
                base_joint = base_item.get('visual_group_joint', base_item.get('Joint', ''))
                base_iso = base_item.get('visual_group_iso', base_item.get('ISO', base_item.get('Dwg', '')))
                
                # Check consecutive items
                for ahead_idx in range(data_ptr + 1, min(len(all_extracted_data), data_ptr + rows_left)):
                    ahead_item = all_extracted_data[ahead_idx]
                    
                    # Verify Joint Merge Configuration
                    ui_merged_j = ahead_item.get('is_merged_joint') == True
                    name_merged_j = ahead_item.get('visual_group_joint', ahead_item.get('Joint', '')) == base_joint
                    joint_ok = ui_merged_j or name_merged_j

                    # Verify ISO Merge Configuration
                    ui_merged_i = ahead_item.get('is_merged_iso') == True
                    name_merged_i = ahead_item.get('visual_group_iso', ahead_item.get('ISO', ahead_item.get('Dwg', ''))) == base_iso
                    iso_ok = ui_merged_i or name_merged_i
                    
                    # If BOTH are ok, it's a block. If EITHER is broken (by User or Data mismatch), we split the block!
                    if joint_ok and iso_ok:
                        logical_block_size += 1
                    else:
                        break
                
                # Cap the block size strictly to the available rows remaining on the sheet
                this_block_size = min(logical_block_size, rows_left)
                actual_block_rows = this_block_size
                batch = all_extracted_data[data_ptr : data_ptr + this_block_size]
                
                # 테두리 및 서식 적용
                d_height = float(self.config.get('ROW_HEIGHT_DATA', 20.55))
                for r_offset in range(actual_block_rows):
                    r = current_row + r_offset
                    rd = ws.row_dimensions[r]; rd.height = d_height
                    for c in range(1, 14):
                        cell = ws.cell(row=r, column=c)
                        l_s = thin_side; r_s = thin_side
                        t_s = medium_side if r == self.config['START_ROW'] else thin_side
                        b_s = medium_side if r == self.config['DATA_END_ROW'] else thin_side
                        
                        if c == 1: l_s = medium_side
                        if c == 13: r_s = medium_side
                        
                        if c <= 5: # A-E 수직 병합 구역 내부 선 제거
                            if 1 < c < 5: r_s = Side(style=None); l_s = Side(style=None)
                            elif c == 1: r_s = Side(style=None)
                            elif c == 5: l_s = Side(style=None)
                            
                            # 가변 블록 크기에 따른 병합 구역 내부 선 처리
                            if actual_block_rows > 1:
                                if r_offset == 0: b_s = Side(style=None)
                                elif r_offset == actual_block_rows - 1: t_s = Side(style=None)
                                else: t_s = Side(style=None); b_s = Side(style=None)
                                
                        cell.border = Border(left=l_s, right=r_s, top=t_s, bottom=b_s)

                # A-E 수평/수직 병합 및 데이터 입력 (블록 전체 범위)
                # [FIX] actual_block_rows가 1이더라도 A~E는 항상 수평 병합되어야 함
                self.safe_merge_cells(ws, start_row=current_row, start_column=1, end_row=current_row + actual_block_rows - 1, end_column=5)
                
                # F열(6)은 2행 이상일 때만 수직 병합
                if actual_block_rows > 1:
                    self.safe_merge_cells(ws, start_row=current_row, start_column=6, end_row=current_row + actual_block_rows - 1, end_column=6)
                
                # 도면번호/조인트번호 입력 (배치의 첫번째 데이터 기준)
                if batch:
                    # [NEW] Dynamic Column Mapping for PMI (Block level)
                    dwg_col = self.col_to_num(self.config.get('PMI_COL_DWG', '1'))
                    joint_col = self.col_to_num(self.config.get('PMI_COL_JOINT', '6'))
                    
                    if dwg_col >= 1:
                        self.safe_set_value(ws, ws.cell(row=current_row, column=dwg_col).coordinate, batch[0].get('Dwg', ''))
                        ws.cell(row=current_row, column=dwg_col).alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                    
                    if joint_col >= 1:
                        self.safe_set_value(ws, ws.cell(row=current_row, column=joint_col).coordinate, batch[0].get('Joint', batch[0].get('No', '')))
                        ws.cell(row=current_row, column=joint_col).alignment = Alignment(horizontal='center', vertical='center')

                # 개별 데이터 기입
                for i, item in enumerate(batch):
                    r = current_row + i
                    loc_col = self.col_to_num(self.config.get('PMI_COL_LOC', '7'))
                    if loc_col >= 1:
                        self.safe_set_value(ws, ws.cell(row=r, column=loc_col).coordinate, item.get('Loc', ''), align='center')
                    
                    for val_key, col_idx in print_elements:
                        if col_idx < 1: continue
                        raw_v = item.get(val_key, 0.0)
                        v = self.to_float(raw_v)
                        cell = ws.cell(row=r, column=col_idx)
                        self.safe_set_value(ws, cell.coordinate, v if v > 0 else "", align='center')
                        if v > 0: cell.number_format = '0.00'
                    
                    res_col = self.col_to_num(self.config.get('PMI_COL_RES', '13'))
                    if res_col >= 1:
                        self.safe_set_value(ws, ws.cell(row=r, column=res_col).coordinate, item.get('Grade', ''), align='center')
                        cell_l = ws.cell(row=r, column=res_col); cell_l.font = Font(size=8.5); dv_q.add(cell_l)

                data_ptr += len(batch)
                current_row += actual_block_rows
                self.progress['value'] = 30 + (data_ptr / len(all_extracted_data)) * 65

            # [NEW] 마지막 데이터 기입 직후 다음 행 H~L에 BLANK 표시 (병합 적용)
            try:
                # current_row는 이미 다음 3행 블록의 시작점이므로, 실제 마지막 데이터 r의 다음 행을 계산
                last_data_r = (current_row - actual_block_rows) + len(batch)
                if last_data_r <= self.config['DATA_END_ROW']:
                    # H(8)~L(12) 열 병합 후 BLANK 입력
                    self.safe_merge_cells(ws, start_row=last_data_r, start_column=8, end_row=last_data_r, end_column=12)
                    self.safe_set_value(ws, ws.cell(row=last_data_r, column=8).coordinate, "BLANK")
                    ws.cell(row=last_data_r, column=8).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=last_data_r, column=8).font = Font(size=9, bold=True)
            except: pass

            # [RE-NC] 데이터가 없는 빈 칸도 45열까지 3개 행 블록 단위로 테두리/병합 적용
            d_height = float(self.config.get('ROW_HEIGHT_DATA', 20.55))
            while current_row <= self.config['DATA_END_ROW']:
                rows_left = self.config['DATA_END_ROW'] - current_row + 1
                this_block_size = min(3, rows_left)
                
                for r_offset in range(this_block_size):
                    r = current_row + r_offset
                    rd = ws.row_dimensions[r]; rd.height = d_height
                    for c in range(1, 14):
                        cell = ws.cell(row=r, column=c)
                        l_s = thin_side; r_s = thin_side
                        t_s = medium_side if r == self.config['START_ROW'] else thin_side
                        b_s = medium_side if r == self.config['DATA_END_ROW'] else thin_side
                        
                        if c == 1: l_s = medium_side
                        if c == 13: r_s = medium_side
                        
                        # [FIX] 기입되지 않은 행이라도 도면번호(A~E) 수평 병합은 유지
                        if 1 <= c <= 5:
                            if 1 < c < 5: r_s = Side(style=None); l_s = Side(style=None)
                            elif c == 1: r_s = Side(style=None)
                            elif c == 5: l_s = Side(style=None)
                            
                        cell.border = Border(left=l_s, right=r_s, top=t_s, bottom=b_s)
                    
                    # 매 행마다 A~E 수평 병합 실행
                    self.safe_merge_cells(ws, start_row=r, start_column=1, end_row=r, end_column=5)
                
                dv_q.add(ws.cell(row=current_row, column=13))
                current_row += this_block_size

            # [FORCE] 모든 데이터 시트의 45행 바닥선(A-L) 설정
            data_end_row = int(self.config.get('DATA_END_ROW', 45))
            for idx, s in enumerate(wb.worksheets):
                if s.max_row >= data_end_row:
                    for c in range(1, 14):
                        cell = s.cell(row=data_end_row, column=c)
                        curr_border = cell.border
                        l_s = curr_border.left; r_s = curr_border.right; t_s = curr_border.top
                        
                        if idx == 0: # 갑지
                            if c in [1, 2, 3, 11, 12, 13]: # A, B, C, K, L, M 바닥선 제거
                                b_s = Side(style=None)
                            else: # D~J 는 약한선
                                b_s = thin_side
                        else: # 을지
                            b_s = medium_side
                            
                        cell.border = Border(left=l_s, right=r_s, top=t_s, bottom=b_s)

            # [FORCE] 을지 시트 상단 외곽 테두리 제거 (1행 Top 및 A1-A4 Left, T1-T4 Right)
            for idx, s in enumerate(wb.worksheets):
                if idx > 0: # 을지 시트
                    # 1행 상단 테두리 제거 (T열까지 확장)
                    for c_idx in range(1, 21):
                        cell = s.cell(row=1, column=c_idx)
                        b = cell.border
                        cell.border = Border(left=b.left, right=b.right, top=Side(style=None), bottom=b.bottom)
                    
                    # A1-A4 좌측, M1-M4 우측, T1-T4 우측 테두리 제거
                    for r_idx in range(1, 5):
                        # A열 (좌측)
                        cell_a = s.cell(row=r_idx, column=1)
                        b_a = cell_a.border
                        cell_a.border = Border(left=Side(style=None), right=b_a.right, top=b_a.top, bottom=b_a.bottom)
                        
                        # M열 (우측)
                        cell_m = s.cell(row=r_idx, column=13)
                        b_m = cell_m.border
                        cell_m.border = Border(left=b_m.left, right=Side(style=None), top=b_m.top, bottom=b_m.bottom)
                        
                        # T열 (우측)
                        cell_t = s.cell(row=r_idx, column=20)
                        b_t = cell_t.border
                        cell_t.border = Border(left=b_t.left, right=Side(style=None), top=b_t.top, bottom=b_t.bottom)
                    
                    # [NEW] L3, M3 (페이지 번호 구역) 모든 테두리 제거
                    for c_idx in [12, 13]: # L, M
                        s.cell(row=3, column=c_idx).border = Border()
                    
                    # [NEW] 5행 위쪽 선 진한선으로 (A~M열 범위)
                    for c_idx in range(1, 14): # A~M
                        cell_5 = s.cell(row=5, column=c_idx)
                        b_5 = cell_5.border
                        cell_5.border = Border(left=b_5.left, right=b_5.right, top=medium_side, bottom=b_5.bottom)

            # [FORCE] 을지 시트의 지정 행 글꼴 크기 통일
            for idx, s in enumerate(wb.worksheets):
                if idx > 0: # 을지 시트
                    for r_idx in range(data_start_row, data_end_row + 1):
                        for c_idx in range(1, 14):
                            cell = s.cell(row=r_idx, column=c_idx)
                            f = cell.font
                            if f:
                                cell.font = Font(name=f.name, size=10, bold=f.bold, italic=f.italic, vertAlign=f.vertAlign, underline=f.underline, strike=f.strike, color=f.color)
                            else:
                                cell.font = Font(name='맑은 고딕', size=10)

            total_p = len(wb.worksheets)
            for p_idx, s in enumerate(wb.worksheets):
                page_num = p_idx + 1
                # [NEW] 최종 저장 직전 사용자의 행/열 커스텀 설정을 모든 시트에 다시 한번 적용 (최우선순위)
                ctx = "COVER" if p_idx == 0 else "DATA"
                self.apply_custom_dimensions(s, ctx)
                
                try:
                    if p_idx == 0: # Gapji (Cover)
                        self.safe_merge_cells(s, 3, 14, 3, 20) # N3:T3
                        cell = s["N3"]
                    else: # Eulji (Data)
                        # [FIX] 을지 시트 페이지 번호 가독성 확보를 위해 L3:M3 병합 검토 (또는 M3 단독)
                        try: self.safe_merge_cells(s, 3, 12, 3, 13) # L3:M3
                        except: pass
                        cell = s["L3"] if "L3" in s.merged_cells else s["M3"]
                    
                    cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                    cell.font = Font(name='맑은 고딕', size=11, bold=True)
                    # 단어 사이의 간격을 인위적으로 늘림
                    self.safe_set_value(s, cell.coordinate, f"Page    {page_num}    of    {total_p}")
                except Exception as e:
                    self.log(f"페이지 번호 기입 실패 ({p_idx}): {e}")

            now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_name = f"{os.path.splitext(os.path.basename(template_path))[0]}_Unified_{now_str}{os.path.splitext(template_path)[1]}"
            save_path = os.path.join(os.path.dirname(template_path), output_name); wb.save(save_path)
            self.progress['value'] = 100; self.log(f"✨ 완료! 저장됨: {output_name}")
            messagebox.showinfo("성공", f"통합 성적서 생성이 완료되었습니다.\n\n경로: {os.path.dirname(save_path)}\n파일명: {output_name}")
        except Exception as e:
            self.log(f"❌ 오류: {e}")
            traceback.print_exc()
            # [NEW] 사용자에게 친절한 팝업 알림 추가
            error_msg = f"성적서 생성 중 오류가 발생했습니다.\n\n원인: {e}\n\n* 주로 양식 파일의 시트 구조나 병합 상태가 맞지 않을 때 발생합니다. 다른 양식을 시도하거나 설정을 확인해주세요."
            messagebox.showerror("생성 실패", error_msg)
            # End of _run_pmi_process logic
        finally:
            if 'wb' in locals() and wb:
                try:
                    if hasattr(wb, 'vba_archive') and wb.vba_archive:
                        wb.vba_archive.close()
                except: pass
                try: wb.close()
                except: pass
            for f in glob.glob(os.path.join(tempfile.gettempdir(), "temp_*.png")):
                try: os.remove(f)
                except: pass

    def _run_pt_process(self, final_list, template_path):
        """PT 성적서 생성 (1-row-per-data 레이아웃 + ISO 병합)"""
        self.log(f"🚀 PT 성적서 생성 시작 (총 {len(final_list)} 건)...")
        self.progress['value'] = 0
        try:
            wb = openpyxl.load_workbook(template_path, keep_vba=True)
            if len(wb.worksheets) < 1:
                raise ValueError("선택한 템플릿 파일에 시트가 존재하지 않습니다.")

            # 갑지 (Cover)
            ws0 = wb.worksheets[0]
            self.add_logos_to_sheet(ws0, is_cover=True, clear_existing=False)
            self.force_print_settings(ws0, context="COVER")
            self.apply_custom_dimensions(ws0, "COVER")

            # 을지 (Data)
            data_sheet_id = 1 if len(wb.worksheets) >= 2 else 0
            ws = wb.worksheets[data_sheet_id]
            ws.title = f"{ws.title[:20]}_001"
            self.add_logos_to_sheet(ws, is_cover=False)
            self.force_print_settings(ws, context="DATA")

            # 헤더 감지 및 시작 행 결정
            start_row = int(self.config.get('PT_START_ROW', 18))
            end_row = int(self.config.get('PT_END_ROW', 37))
            
            # 스타일 설정
            thin_side = Side(style='thin')
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            # [Smart Detection] 템플릿에서 헤더를 찾아 시작 행 자동 보정
            detected_start = -1
            iso_col_idx = 2 # 기본값 B컬럼
            for r in range(1, 40):
                line_cells = [str(ws.cell(row=r, column=c).value).upper() for c in range(1, 15)]
                line_str = "".join(line_cells)
                if (("JOINT" in line_str or "WELD" in line_str) and ("ISO" in line_str or "LINE" in line_str)) or ("DWG" in line_str and "RESULT" in line_str):
                    detected_start = r + 1
                    # ISO 컬럼 위치 동적 확인
                    for c_idx, val in enumerate(line_cells):
                        if any(k in val for k in ["ISO", "LINE", "DWG", "DRAWING"]) and not any(k in val for k in ["JOINT", "WELD"]):
                            iso_col_idx = c_idx + 1
                            break
                    break
            if detected_start > 0: start_row = detected_start

            current_row = start_row
            current_page = 1
            data_ptr = 0
            
            while data_ptr < len(final_list):
                if current_row > end_row:
                    current_page += 1
                    ws = self.prepare_next_sheet(wb, data_sheet_id, current_page)
                    current_row = start_row
                
                item = final_list[data_ptr]
                
                # 데이터 기입 (B, C, D 컬럼 병합 대응)
                no_col = self.col_to_num(self.config.get('PT_COL_NO', '1'))
                if no_col >= 1:
                    self.safe_set_value(ws, ws.cell(row=current_row, column=no_col).coordinate, item.get('No', ''))
                
                # ISO Drawing No. (병합된 칸 반영)
                dwg_col = self.col_to_num(self.config.get('PT_COL_DWG', str(iso_col_idx)))
                if dwg_col >= 1:
                    iso_cell = ws.cell(row=current_row, column=dwg_col)
                    self.safe_set_value(ws, iso_cell.coordinate, item.get('Dwg', ''))
                    # Preserve original 3-column merge if using default or close to it
                    if dwg_col == iso_col_idx:
                        self.safe_merge_cells(ws, current_row, dwg_col, current_row, dwg_col + 2)
                
                joint_col = self.col_to_num(self.config.get('PT_COL_JOINT', '5'))
                if joint_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=joint_col).coordinate, item.get('Joint', ''))
                
                nps_col = self.col_to_num(self.config.get('PT_COL_NPS', '6'))
                if nps_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=nps_col).coordinate, item.get('NPS', ''))
                
                thk_col = self.col_to_num(self.config.get('PT_COL_THK', '7'))
                if thk_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=thk_col).coordinate, item.get('Thk.', ''))
                
                mat_col = self.col_to_num(self.config.get('PT_COL_MAT', '8'))
                if mat_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=mat_col).coordinate, item.get('Material', ''))
                
                wld_col = self.col_to_num(self.config.get('PT_COL_WELDER', '9'))
                if wld_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=wld_col).coordinate, item.get('Welder', ''))
                
                typ_col = self.col_to_num(self.config.get('PT_COL_TYPE', '10'))
                if typ_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=typ_col).coordinate, item.get('WType', ''))
                
                res_col = self.col_to_num(self.config.get('PT_COL_RES', '11'))
                if res_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=res_col).coordinate, item.get('Result', 'Acc'))
                
                # 스타일링
                for c in range(1, 12):
                    cell = ws.cell(row=current_row, column=c)
                    cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                    cell.font = Font(name='바탕', size=9)
                    cell.border = thin_border

                data_ptr += 1
                current_row += 1
                self.progress['value'] = (data_ptr / len(final_list)) * 95

            total_p = len(wb.worksheets)
            for p_idx, s in enumerate(wb.worksheets):
                page_num = p_idx + 1
                self.apply_custom_dimensions(s, "DATA" if p_idx > 0 else "COVER")
                # 페이지 번호 기입
                try:
                    p_text = f"Page    {page_num}    of    {total_p}"
                    if p_idx == 0: self.safe_set_value(s, 'O35', p_text)
                    else: self.safe_set_value(s, 'V3', p_text)
                except: pass

            now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_name = f"PT_Report_{now_str}.xlsx"
            save_path = os.path.join(os.path.dirname(template_path), output_name)
            wb.save(save_path)
            self.progress['value'] = 100
            self.log(f"✨ PT 완료! 저장됨: {output_name}")
            messagebox.showinfo("성공", f"PT 성적서 생성이 완료되었습니다.\n경로: {os.path.dirname(save_path)}")
        except Exception as e:
            self.log(f"❌ PT 생성 오류: {e}")
            traceback.print_exc()
            messagebox.showerror("생성 실패", f"PT 성적서 생성 중 오류가 발생했습니다: {e}")
        finally:
            if 'wb' in locals() and wb:
                try:
                    if hasattr(wb, 'vba_archive') and wb.vba_archive:
                        wb.vba_archive.close()
                except: pass
                try: wb.close()
                except: pass
            for f in glob.glob(os.path.join(tempfile.gettempdir(), "temp_*.png")):
                try: os.remove(f)
                except: pass

    def _run_rt_process(self, final_list, template_path):
        """RT 성적서 생성 (1-row-per-data 레이아웃)"""
        self.log(f"🚀 RT 성적서 생성 시작 (총 {len(final_list)} 건)...")
        self.progress['value'] = 0
        try:
            wb = openpyxl.load_workbook(template_path, keep_vba=True)
            if len(wb.worksheets) < 1:
                raise ValueError("선택한 템플릿 파일에 시트가 존재하지 않습니다.")

            # Gapji (Cover) Logic (Simplified for RT, using generic SITCO style)
            ws0 = wb.worksheets[0]
            self.add_logos_to_sheet(ws0, is_cover=True, clear_existing=False)
            self.force_print_settings(ws0, context="COVER")
            self.apply_custom_dimensions(ws0, "COVER")

            # Data Sheet Logic (1-row-per-data)
            data_sheet_id = 1 if len(wb.worksheets) >= 2 else 0
            ws = wb.worksheets[data_sheet_id]
            ws.title = f"{ws.title[:20]}_001"
            self.add_logos_to_sheet(ws, is_cover=False)
            self.force_print_settings(ws, context="DATA")
            # RT usually doesn't have the same headers as PMI
            # self.set_eulji_headers(ws) 

            # Configizable or Default RT boundaries
            start_row = int(self.config.get('RT_START_ROW', 17))
            end_row = int(self.config.get('RT_END_ROW', 41))
            rows_per_page = end_row - start_row + 1
            
            current_row = start_row
            current_page = 1
            data_ptr = 0
            
            while data_ptr < len(final_list):
                if current_row > end_row:
                    current_page += 1
                    ws = self.prepare_next_sheet(wb, data_sheet_id, current_page)
                    current_row = start_row
                
                item = final_list[data_ptr]
                
                # [NEW] Write headers to Excel on the row above start_row
                if current_row == start_row:
                    try:
                        h_row = start_row - 1
                        if h_row >= 1:
                            # Use existing mapping keys to place headers
                            h_map = {
                                'RT_COL_NO': 'RT_NAME_NO', 'RT_COL_DATE': 'RT_NAME_DATE',
                                'RT_COL_DWG': 'RT_NAME_DWG', 'RT_COL_JOINT': 'RT_NAME_JOINT',
                                'RT_COL_LOC': 'RT_NAME_LOC', 'RT_COL_THK': 'RT_NAME_THK',
                                'RT_COL_MAT': 'RT_NAME_MAT', 'RT_COL_ACC': 'RT_NAME_ACC',
                                'RT_COL_REJ': 'RT_NAME_REJ', 'RT_COL_DEG': 'RT_NAME_DEG',
                                'RT_COL_RES': 'RT_NAME_RES', 'RT_COL_WELDER': 'RT_NAME_WELDER',
                                'RT_COL_REM': 'RT_NAME_REM'
                            }
                            for c_key, n_key in h_map.items():
                                c_idx = self.col_to_num(self.config.get(c_key, '0'))
                                if c_idx >= 1:
                                    h_default = c_key.replace('RT_COL_', '').capitalize()
                                    h_text = self.config.get(n_key, h_default)
                                    self.safe_set_value(ws, ws.cell(row=h_row, column=c_idx).coordinate, h_text)
                            
                            # Defects D1-D15 headers
                            for d_i in range(1, 16):
                                d_c_idx = self.col_to_num(self.config.get(f'RT_COL_D{d_i}', '0'))
                                if d_c_idx >= 1:
                                    d_text = self.config.get(f'RT_NAME_D{d_i}', f'D{d_i}')
                                    self.safe_set_value(ws, ws.cell(row=h_row, column=d_c_idx).coordinate, d_text)
                    except: pass

                # Column Data Mapping (standardized layout)
                no_col = self.col_to_num(self.config.get('RT_COL_NO', '1'))
                if no_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=no_col).coordinate, item.get('No', ''))
                
                date_col = self.col_to_num(self.config.get('RT_COL_DATE', '2'))
                if date_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=date_col).coordinate, item.get('Date', ''))
                
                dwg_col = self.col_to_num(self.config.get('RT_COL_DWG', '3'))
                if dwg_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=dwg_col).coordinate, item.get('Dwg', ''))
                
                joint_col = self.col_to_num(self.config.get('RT_COL_JOINT', '4'))
                if joint_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=joint_col).coordinate, item.get('Joint', ''))
                
                loc_col = self.col_to_num(self.config.get('RT_COL_LOC', '5'))
                if loc_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=loc_col).coordinate, item.get('Loc', ''))
                
                # Standard columns (Dynamic mapping)
                thk_col = self.col_to_num(self.config.get('RT_COL_THK', '6'))
                if thk_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=thk_col).coordinate, item.get('T', ''))
                
                mat_col = self.col_to_num(self.config.get('RT_COL_MAT', '7'))
                if mat_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=mat_col).coordinate, item.get('Mat', ''))
                
                acc_col = self.col_to_num(self.config.get('RT_COL_ACC', '7'))
                if acc_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=acc_col).coordinate, item.get('Acc', ''))
                
                rej_col = self.col_to_num(self.config.get('RT_COL_REJ', '8'))
                if rej_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=rej_col).coordinate, item.get('Rej', ''))
                
                deg_col = self.col_to_num(self.config.get('RT_COL_DEG', '9'))
                if deg_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=deg_col).coordinate, item.get('Deg', ''))

                # IQI/Sens/Den (Legacy or fixed for now)
                iqi_col = self.col_to_num(self.config.get('RT_COL_IQI', '10'))
                if iqi_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=iqi_col).coordinate, item.get('IQI', ''))
                
                sens_col = self.col_to_num(self.config.get('RT_COL_SENS', '11'))
                if sens_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=sens_col).coordinate, item.get('Sens', ''))
                
                den_col = self.col_to_num(self.config.get('RT_COL_DEN', '12'))
                if den_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=den_col).coordinate, item.get('Den', ''))
                
                # Defects D1-D15 (Dynamic)
                for d_i in range(1, 16):
                    d_col_key = f'RT_COL_D{d_i}'
                    d_col_idx = self.col_to_num(self.config.get(d_col_key, str(12 + d_i)))
                    if d_col_idx >= 1:
                        self.safe_set_value(ws, ws.cell(row=current_row, column=d_col_idx).coordinate, item.get(f'D{d_i}', ''))
                
                res_col = self.col_to_num(self.config.get('RT_COL_RES', '28'))
                if res_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=res_col).coordinate, item.get('Result', 'ACC'))
                
                wld_col = self.col_to_num(self.config.get('RT_COL_WELDER', '29'))
                if wld_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=wld_col).coordinate, item.get('Welder', ''))
                
                rem_col = self.col_to_num(self.config.get('RT_COL_REM', '30'))
                if rem_col >= 1: self.safe_set_value(ws, ws.cell(row=current_row, column=rem_col).coordinate, item.get('Remarks', ''))
                
                # Styling for individual row
                for c in range(1, 31):
                    cell = ws.cell(row=current_row, column=c)
                    cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                    cell.font = Font(name='바탕', size=9)

                data_ptr += 1
                current_row += 1
                self.progress['value'] = (data_ptr / len(final_list)) * 95

            total_p = len(wb.worksheets)
            for p_idx, s in enumerate(wb.worksheets):
                page_num = p_idx + 1
                self.apply_custom_dimensions(s, "DATA" if p_idx > 0 else "COVER")
                # Page Numbers
                try:
                    if p_idx == 0: cell = s["O35"] if "O35" in s else s.cell(row=35, column=15)
                    else: cell = s["V3"] if "V3" in s else s.cell(row=3, column=22)
                    self.safe_set_value(s, cell.coordinate, f"Page   {page_num}   of   {total_p}")
                except: pass

            now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_name = f"RT_Report_{now_str}.xlsx"
            save_path = os.path.join(os.path.dirname(template_path), output_name)
            wb.save(save_path)
            self.progress['value'] = 100
            self.log(f"✨ RT 완료! 저장됨: {output_name}")
            messagebox.showinfo("성공", f"RT 성적서 생성이 완료되었습니다.\n경로: {os.path.dirname(save_path)}")
        except Exception as e:
            self.log(f"❌ RT 생성 오류: {e}")
            traceback.print_exc()
            messagebox.showerror("생성 실패", f"RT 성적서 생성 중 오류가 발생했습니다: {e}")
        finally:
            if 'wb' in locals() and wb:
                try:
                    if hasattr(wb, 'vba_archive') and wb.vba_archive:
                        wb.vba_archive.close()
                except: pass
                try: wb.close()
                except: pass
            for f in glob.glob(os.path.join(tempfile.gettempdir(), "temp_*.png")):
                try: os.remove(f)
                except: pass

    def _on_entry_esc(self, event):
        """Removes focus and clears selection on ESC (preserves text)."""
        try:
            widget = event.widget
            if hasattr(widget, 'selection_clear'):
                widget.selection_clear()
            self.root.focus_set()
        except: pass

if __name__ == "__main__":
    root = tk.Tk()
    PMIReportApp(root)
    root.mainloop()
