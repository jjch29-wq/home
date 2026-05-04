import openpyxl
import os

path = r"c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\Na-aba\home\src\Template_DailyWorkReport.xlsx"
if not os.path.exists(path):
    print(f"File not found: {path}")
else:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active
    for r in range(20, 60):
        val_a = sheet.cell(row=r, column=1).value
        val_b = sheet.cell(row=r, column=2).value
        if val_a or val_b:
            print(f"Row {r}: A={val_a}, B={val_b}")
