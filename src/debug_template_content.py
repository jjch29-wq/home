import openpyxl
import os

path = r'c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\Na-aba\home\resources\Template_DailyWorkReport.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active
val = ws.cell(row=27, column=1).value
print(f"Content: {val}")
if val:
    print(f"Hex: {val.encode('utf-8').hex()}")
else:
    print("Content is None")

# Also check row 26 just in case
val26 = ws.cell(row=26, column=1).value
print(f"Row 26: {val26}")
if val26:
    print(f"Row 26 Hex: {val26.encode('utf-8').hex()}")
