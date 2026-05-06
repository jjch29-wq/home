from openpyxl import load_workbook
import os

path = os.path.abspath("../resources/Template_DailyWorkReport.xlsx")
wb = load_workbook(path)
sheet = wb.active

for r in range(1, 100):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 20)]
    if any(x and ("품명" in str(x) or "규격" in str(x)) for x in row_vals):
        print(f"Header Row {r}: {row_vals}")

wb.close()
