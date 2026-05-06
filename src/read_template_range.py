import openpyxl
import os

def read_cell_range(file_path, start_row, end_row, start_col, end_col):
    if not os.path.exists(file_path):
        return f"File not found: {file_path}"
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        results = []
        for r in range(start_row, end_row + 1):
            row_vals = []
            for c in range(start_col, end_col + 1):
                col_letter = openpyxl.utils.get_column_letter(c)
                val = sheet.cell(row=r, column=c).value
                row_vals.append(f"{col_letter}{r}: {val}")
            results.append(" | ".join(row_vals))
        return "\n".join(results)
    except Exception as e:
        return f"Error: {e}"

current_dir = r"c:\Users\jjch2\Desktop\보고서\Project PROVIDENCE\Request\PMI\Na-aba\home\src"
home_dir = os.path.dirname(current_dir)
template_path = os.path.join(home_dir, 'resources', 'Template_DailyWorkReport.xlsx')

print(f"Reading range in {template_path}...")
print(read_cell_range(template_path, 20, 35, 1, 16)) # A20 to P35
