import openpyxl
from openpyxl.styles import Border, Side

wb = openpyxl.load_workbook(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')

thin_side = Side(style='thin')
hair_side = Side(style='hair')

for sheet_name in ['PAUT', 'PT', 'MT']:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Headers are usually on row 1, data starts at row 2
        min_r = 2
        max_r = ws.max_row
        min_c = 1
        max_c = ws.max_column
        
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                cell = ws.cell(row=r, column=c)
                
                # Keep inner borders as hair, but set outer to thin
                left = hair_side
                right = hair_side
                top = hair_side
                bottom = hair_side
                
                if c == min_c: left = thin_side
                if c == max_c: right = thin_side
                if r == min_r: top = thin_side
                if r == max_r: bottom = thin_side
                
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

wb.save(r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx')
print("Outer borders of PAUT, PT, MT updated to solid thin.")
