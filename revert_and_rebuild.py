import shutil
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter

# 1. Start from the original preserved format file
src_path = r'C:\Users\jjch2\Desktop\중앙지사 paut 리스트_서식포함.xlsx'
dst_path = r'C:\Users\jjch2\Desktop\중앙지사 종합 리스트.xlsx'
shutil.copy(src_path, dst_path)

wb = openpyxl.load_workbook(dst_path)
source_sheet = wb.active
source_sheet.title = 'PAUT'

# 2. Create PT and MT in the old format
methods = ['PT', 'MT']
for m in methods:
    new_ws = wb.copy_worksheet(source_sheet)
    new_ws.title = m
    for r in new_ws.iter_rows():
        for c in r:
            if isinstance(c.value, str):
                c.value = c.value.replace('PAUT', m).replace('paut', m.lower())

# 3. Create the new RT sheet
if 'RT' in wb.sheetnames:
    del wb['RT']
ws_rt = wb.create_sheet('RT')

# RT Headers Row 1 and 2
row1 = ['순번', '제조사', '구분', '촬영\n일자', 'Section', 'Line No.', 'Joint', '용접사\n번호', 
        '촬영구간', '', '', '', '', '', '', '', '관경', '필름규격', '장수', 'R', '촬영결과', '비고']
ws_rt.append(row1)
row2 = ['', '', '', '', '', '', '', '', '1', '2', '3', '4', '5', '6', '7', '8', '', '', '', '', '', '']
ws_rt.append(row2)

# RT Merges
merges = ['A1:A2', 'B1:B2', 'C1:C2', 'D1:D2', 'E1:E2', 'F1:F2', 'G1:G2', 'H1:H2',
          'I1:P1', 'Q1:Q2', 'R1:R2', 'S1:S2', 'T1:T2', 'U1:U2', 'V1:V2']
for m in merges:
    ws_rt.merge_cells(m)

# RT Styling (Headers)
fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
font_bold = Font(bold=True)
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

for row in ws_rt.iter_rows(min_row=1, max_row=2, min_col=1, max_col=22):
    for cell in row:
        cell.fill = fill
        cell.border = border_thin
        cell.font = font_bold
        cell.alignment = align_center

# RT Column widths
widths = [6, 10, 6, 12, 12, 30, 8, 10] + [5]*8 + [10, 15, 6, 6, 10, 15]
for i, w in enumerate(widths, 1):
    ws_rt.column_dimensions[get_column_letter(i)].width = w

# RT Dummy data
dummy_data = [
    [1, '세경', '', '21.09.11', 'SEC 4', 'HC B 112 300A', 'R01', 'WD04', '1', '1', '1', '1', '', '', '', '', '300A', '3 1/3 x 12"', 4, '', '합격', ''],
    [2, '세경', '', '21.09.11', 'SEC 4', 'HC B 112 300A', 'R02', 'WD04', 'P/1', '1', '1', '1', '', '', '', '', '300A', '3 1/3 x 12"', 4, '', '합격', ''],
    [3, '세경', '', '21.09.11', 'SEC 4', 'HC B 112 300A', 'R03', 'WD04', 'P/1', '1', '1', '1', '', '', '', '', '300A', '3 1/3 x 12"', 4, '', '합격', '']
]
for row_data in dummy_data:
    ws_rt.append(row_data)
    for cell in ws_rt[ws_rt.max_row]:
        cell.alignment = align_center

# RT Borders (Data rows)
dashed_side = Side(style='dashed')
thin_side = Side(style='thin')
for r in range(3, ws_rt.max_row + 1):
    for c in range(1, 23):
        cell = ws_rt.cell(row=r, column=c)
        left = thin_side if c == 1 else dashed_side
        right = thin_side if c == 22 else dashed_side
        top = thin_side if r == 3 else dashed_side
        bottom = thin_side if r == ws_rt.max_row else dashed_side
        cell.border = Border(left=left, right=right, top=top, bottom=bottom)

ws_rt.auto_filter.ref = f'A2:V{ws_rt.max_row}'

# Reorder sheets to: PAUT, RT, PT, MT
# Currently they are PAUT, PT, MT, RT (since RT was created last)
wb.move_sheet(ws_rt, offset=-2)

wb.save(dst_path)
print("Reverted PAUT/PT/MT to old format and preserved RT.")
