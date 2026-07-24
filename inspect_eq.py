import openpyxl

for f in [r'C:\Users\-\PMI\home\src\templates\양식_동탄.xlsx', r'C:\Users\-\PMI\home\src\templates\양식_기본.xlsx']:
    try:
        wb = openpyxl.load_workbook(f)
        for s in wb.sheetnames:
            if "장비" in s:
                print(f'\n--- {f} - {s} ---')
                ws = wb[s]
                for r in ws.iter_rows(values_only=True, max_row=20):
                    row_str = ' | '.join(str(c) for c in r if c and str(c).strip())
                    if row_str:
                        print(row_str)
    except Exception as e:
        print(f"Error on {f}: {e}")
