import openpyxl
import codecs

# Use the ORIGINAL template (not the debug file which already has data written)
wb = openpyxl.load_workbook(r'C:/Users/jjch2/Desktop/템플릿_최종완성본_V74.xlsx', data_only=True)
ws = wb.worksheets[0]

out = codecs.open('section21_detail.txt', 'w', 'utf-8')

# Check all merged cell anchors in area 521-540
out.write("=== Row 521-545 전체 셀 값 (병합 앵커 포함) ===\n")
for row in range(521, 546):
    line = f"Row {row}: "
    has_val = False
    for col in range(1, 30):
        c = ws.cell(row=row, column=col)
        if c.value is not None:
            v = str(c.value).strip().replace('\n', ' ')
            if v:
                line += f"C{col}={repr(v)} "
                has_val = True
    if has_val:
        out.write(line + "\n")

out.write("\n=== 병합 범위 (row 521-545) ===\n")
for merge in ws.merged_cells.ranges:
    if 521 <= merge.min_row <= 545 or 521 <= merge.max_row <= 545:
        out.write(f"  {merge.coord}\n")

out.close()
print("Done! Check section21_detail.txt")
