import openpyxl

wb = openpyxl.load_workbook('home/src/templates/양식_기본.xlsx')
for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    for row in range(1, 1500):
        for col in range(1, 15):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str) and ('SEC' in val or '사진' in val):
                print(f"[{sheetname}] Found at row {row}, col {col}: {val}")
