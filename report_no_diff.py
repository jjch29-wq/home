import json,re,sys
from openpyxl import load_workbook

def n(v):
    if v is None:return ''
    if isinstance(v,float) and v.is_integer():v=int(v)
    return re.sub(r'\s+','',str(v)).upper()
def j(v):
    s=n(v); m=re.fullmatch(r'0*(\d+)(.*)',s)
    return str(int(m.group(1)))+m.group(2) if m else s
def rows(p,s):
    ws=load_workbook(p,data_only=True)[s]; out=[]
    for r in range(5,ws.max_row+1):
        rep=n(ws.cell(r,2).value); m=re.search(r'-RT-(\d+)',rep)
        if not m or int(m.group(1))<94:continue
        out.append({'row':r,'report':rep,'sno':n(ws.cell(r,3).value),'ident':n(ws.cell(r,4).value),'joint':j(ws.cell(r,7).value),'film':ws.cell(r,12).value or 0,'acc':ws.cell(r,13).value or 0,'rej':ws.cell(r,14).value or 0})
    return out
a=rows(sys.argv[1],'정호이엔씨 RT');b=rows(sys.argv[2],'RT List')
ka={(x['report'],x['ident'],x['joint']) for x in a};kb={(x['report'],x['ident'],x['joint']) for x in b}
print(json.dumps({'only_status':[x for x in b if (x['report'],x['ident'],x['joint']) not in ka],'only_jhc':[x for x in a if (x['report'],x['ident'],x['joint']) not in kb]},ensure_ascii=False,indent=2))
