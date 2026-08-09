import openpyxl

wb = openpyxl.load_workbook(r'C:/Users/jjch2/Desktop/템플릿_최종완성본_V74.xlsx', data_only=True)
ws = wb.worksheets[0]

# Check rows around 전체 RT/PAUT/NDT 결과서 부분
# Looking for the actual section with 업체/검사방법/구간/라인번호/Joint No.
print("=== 전체 시트에서 '업체'가 있는 행들 ===")
for row in ws.iter_rows():
    row_vals = {}
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            cv = cell.value.strip().replace('\n', ' ')
            if cv:
                row_vals[cell.column] = cv

    # '업체' 가 있고 검사 관련 헤더가 있는 행
    if any('업체' in v for v in row_vals.values()):
        row_num = list(row)[0].row
        vals_ascii = {k: v.encode('ascii', 'ignore').decode('ascii') for k, v in row_vals.items()}
        print(f"Row {row_num}: {vals_ascii}")

print()

# Also check rows 340-370 (saw that PAUT/RT area was around there)
print("=== Row 355-375 상세 ===")
for row in range(355, 375):
    cells = {}
    for col in range(1, 30):
        c = ws.cell(row=row, column=col)
        if c.value and isinstance(c.value, str):
            cv = c.value.strip().replace('\n', ' ')
            if cv:
                cells[col] = cv.encode('ascii', 'ignore').decode('ascii')
    if cells:
        print(f"Row {row}: {cells}")
